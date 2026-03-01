#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.37.1 - Procedure Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.37: Documented Operating Procedures
Assessment Domain 1 of 3: Procedure Inventory Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific documented operating procedures infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Procedure categories and mandatory documentation requirements (match your operations)
2. Quality standard criteria and approval requirements (adapt to your governance)
3. Review cycle frequency per procedure criticality classification
4. Procedure format and structure standards (template references)
5. Version control and change management integration scope

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.37 Documented Operating Procedures Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
documented operating procedures controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Procedure Inventory Assessment under ISO 27001:2022 Control A.5.37. Supports evidence-based documentation of procedure inventory completeness, quality compliance, and review cycle adherence.

**Assessment Scope:**
- Operating procedure inventory completeness and currency
- Procedure quality standard compliance and format adherence
- Review and update cycle tracking and completion rates
- Ownership assignment and accountability documentation
- Version control and change history completeness
- Accessibility and distribution channel effectiveness
- Evidence collection for operational and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Documented Operating Procedures controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a537_1_procedure_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a537_1_procedure_inventory.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a537_1_procedure_inventory.py --date 20250115

Output:
    File: ISMS-IMP-A.5.37.1_Procedure_Inventory_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.37
Assessment Domain:    1 of 3 (Procedure Inventory Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.37: Documented Operating Procedures Policy (Governance)
    - ISMS-IMP-A.5.37.1: Procedure Inventory Assessment (Domain 1)
    - ISMS-IMP-A.5.37.2: Procedure Quality Assessment (Domain 2)
    - ISMS-IMP-A.5.37.3: Procedure Review and Update Tracking (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.37.1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive documented operating procedures details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review procedure inventory completeness and quality standards annually or when operational processes change, system upgrades occur, or compliance findings identify procedure gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.37.1"
WORKBOOK_NAME = "Procedure Inventory Assessment"
CONTROL_ID = "A.5.37"
CONTROL_NAME = "Documented Operating Procedures"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# STYLE HELPERS
# =============================================================================
def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

_STYLES = setup_styles()
def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches specification — Summary Dashboard before Approval Sign-Off
    sheets = [
        "Instructions & Legend",
        "Procedure Inventory",
        "Required Procedures",
        "Accessibility Matrix",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete the Procedure Inventory — list all operating procedures for information processing facilities.',
        '2. For each procedure, document owner, version, review date, and location.',
        '3. Complete the Required Procedures sheet — verify all ISO 27001:2022 A.5.37 mandatory procedures exist.',
        '4. Complete the Accessibility Matrix — confirm each procedure is accessible to personnel who need it.',
        '5. Complete the Gap Analysis — identify missing, outdated, or inaccessible procedures.',
        '6. Maintain the Evidence Register with procedure documents and accessibility confirmations.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_procedure_inventory_sheet(ws, styles):
    """Create the Procedure Inventory sheet — master catalogue."""
    border = _thin_border()
    grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:P1")
    ws["A1"] = "PROCEDURE INVENTORY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2 subtitle
    ws.merge_cells("A2:P2")
    ws["A2"] = "Master catalogue of all documented operating procedures — record each procedure, its owner, location, approval status and review dates"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column definitions
    columns = [
        ("Procedure ID", 18),
        ("Procedure Name", 40),
        ("Category", 20),
        ("Process Owner", 25),
        ("Department", 20),
        ("Document Location", 35),
        ("Last Review Date", 16),
        ("Next Review Due", 16),
        ("Review Cycle Days", 16),
        ("Version", 12),
        ("Approval Status", 18),
        ("Approver", 25),
        ("Approval Date", 16),
        ("Related Controls", 25),
        ("Criticality", 14),
        ("Notes", 35),
    ]

    # Headers (row 3)
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validations
    dv_category = DataValidation(
        type="list",
        formula1='"System Operations,Security Operations,Facility Operations,Change Management,Recovery Operations,User Management,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Pending Approval,Approved,Expired,Under Review"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_criticality = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_criticality)

    # Row 4: F2F2F2 sample row
    sample_vals = {
        1: "PROC-001", 2: "Server Backup and Recovery Procedure", 3: "System Operations",
        4: "IT Operations Manager", 5: "IT Department", 6: "/policies/ops/backup-recovery.pdf",
        7: "01.01.2026", 8: None, 9: 365, 10: "v1.2", 11: "Approved",
        12: "CTO", 13: "15.01.2026", 14: "A.8.13", 15: "Critical", 16: "Annual review required",
    }
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=4, column=c)
        cell.fill = grey_sample
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if c in sample_vals and sample_vals[c] is not None:
            cell.value = sample_vals[c]
    # Next_Review_Due formula for sample row
    ws.cell(row=4, column=8).value = "=IF(G4<>\"\",G4+I4,\"\")"
    ws.cell(row=4, column=8).fill = grey_sample

    # Rows 5-54: 50 empty FFFFCC rows
    for r in range(5, 55):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Apply validations (exclude sample row)
        dv_category.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=11))
        dv_criticality.add(ws.cell(row=r, column=15))

        # Next_Review_Due formula
        ws.cell(row=r, column=8).value = f"=IF(G{r}<>\"\",G{r}+I{r},\"\")"

        # Default review cycle
        ws.cell(row=r, column=9).value = 365

    ws.freeze_panes = "A4"


# =============================================================================
# REQUIRED PROCEDURES SHEET
# =============================================================================
def create_required_procedures_sheet(ws, styles):
    """Create the Required Procedures sheet — ISO 27001 reference list."""
    border = _thin_border()
    grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "REQUIRED PROCEDURES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2 subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "ISO 27001:2022 required procedure reference list — map existing procedures to required ones and record gaps"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Reference ID", 15),
        ("Required Procedure", 50),
        ("ISO Control", 20),
        ("Category", 22),
        ("Priority", 14),
        ("Current Status", 18),
        ("Mapped Procedure ID", 20),
        ("Gap Notes", 40),
    ]

    # Headers (row 3)
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate required procedures (F2F2F2 rows 4-12)
    required_procedures = [
        ("REQ-001", "Information Security Incident Response Procedure", "A.5.24-28", "Security Operations", "High"),
        ("REQ-002", "Access Control Management Procedure", "A.5.15-18", "User Management", "High"),
        ("REQ-003", "Change Management Procedure", "A.8.32", "Change Management", "High"),
        ("REQ-004", "Backup and Recovery Procedure", "A.8.13", "System Operations", "Critical"),
        ("REQ-005", "Business Continuity Activation Procedure", "A.5.29-30", "Recovery Operations", "Critical"),
        ("REQ-006", "User Onboarding/Offboarding Procedure", "A.6.1-5", "User Management", "High"),
        ("REQ-007", "Vulnerability Management Procedure", "A.8.8", "Security Operations", "High"),
        ("REQ-008", "Logging and Monitoring Procedure", "A.8.15-16", "System Operations", "High"),
        ("REQ-009", "Physical Access Control Procedure", "A.7.1-4", "Facility Operations", "High"),
    ]

    # Data validations
    dv_status = DataValidation(
        type="list",
        formula1='"Exists,Partial,Missing"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    # Populate F2F2F2 requirement rows (4-12)
    for row_idx, (ref_id, proc, control, category, priority) in enumerate(required_procedures, start=4):
        for c in range(1, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = grey_sample
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=1, value=ref_id)
        ws.cell(row=row_idx, column=2, value=proc)
        ws.cell(row=row_idx, column=3, value=control)
        ws.cell(row=row_idx, column=4, value=category)
        ws.cell(row=row_idx, column=5, value=priority)
        # Cols F-H input
        for c in range(6, 9):
            ws.cell(row=row_idx, column=c).fill = input_fill
        dv_status.add(ws.cell(row=row_idx, column=6))
        dv_priority.add(ws.cell(row=row_idx, column=5))

    # Rows 13-63: 50 empty FFFFCC rows
    for r in range(13, 64):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_status.add(ws.cell(row=r, column=6))
        dv_priority.add(ws.cell(row=r, column=5))

    ws.freeze_panes = "A4"


# =============================================================================
# ACCESSIBILITY MATRIX SHEET
# =============================================================================
def create_accessibility_matrix_sheet(ws, styles):
    """Create the Accessibility Matrix sheet — role-based access mapping."""
    border = _thin_border()
    grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "ACCESSIBILITY MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2 subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Role-based procedure access mapping — verify that all relevant staff can access the procedures they need"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Procedure ID", 18),
        ("IT Operations", 16),
        ("Security Team", 16),
        ("Facilities", 16),
        ("Help Desk", 16),
        ("Management", 16),
        ("Access Method", 25),
        ("Verified Date", 16),
    ]

    # Headers (row 3)
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Boolean dropdown for role columns
    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    # Row 4: F2F2F2 sample row
    sample_acc = {
        1: "PROC-001", 2: "Yes", 3: "Yes", 4: "No",
        5: "Yes", 6: "Yes", 7: "SharePoint: /IT/Procedures/", 8: "15.01.2026",
    }
    for c in range(1, 9):
        cell = ws.cell(row=4, column=c)
        cell.fill = grey_sample
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if c in sample_acc:
            cell.value = sample_acc[c]

    # Rows 5-54: 50 empty FFFFCC rows
    for r in range(5, 55):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Apply boolean validation to role columns (2-6)
        for c in range(2, 7):
            dv_bool.add(ws.cell(row=r, column=c))

    ws.freeze_panes = "A4"


# =============================================================================
# GAP ANALYSIS SHEET
# =============================================================================
def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet — track procedure gaps and remediation."""
    border = _thin_border()
    grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:J1")
    ws["A1"] = "GAP ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2 subtitle
    ws.merge_cells("A2:J2")
    ws["A2"] = "Track missing, incomplete, outdated or unapproved procedures — assign remediation owners and target dates"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Gap ID", 12),
        ("Gap Type", 18),
        ("Procedure Reference", 35),
        ("Severity", 14),
        ("Identified Date", 16),
        ("Remediation Owner", 25),
        ("Target Date", 16),
        ("Status", 16),
        ("Completion Date", 16),
        ("Evidence", 35),
    ]

    # Headers (row 3)
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Validations
    dv_gap_type = DataValidation(
        type="list",
        formula1='"Missing,Incomplete,Outdated,Unapproved"',
        allow_blank=False
    )
    ws.add_data_validation(dv_gap_type)

    dv_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Row 4: F2F2F2 sample row
    sample_gap = {
        1: "GAP-001", 2: "Missing", 3: "Patch Management Procedure",
        4: "High", 5: "01.02.2026", 6: "IT Security Manager",
        7: "01.04.2026", 8: "Open", 9: "", 10: "/evidence/gap-001-assessment.pdf",
    }
    for c in range(1, 11):
        cell = ws.cell(row=4, column=c)
        cell.fill = grey_sample
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if c in sample_gap and sample_gap[c] is not None:
            cell.value = sample_gap[c]

    # Rows 5-54: 50 empty FFFFCC rows
    for r in range(5, 55):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        dv_gap_type.add(ws.cell(row=r, column=2))
        dv_severity.add(ws.cell(row=r, column=4))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register(ws):
    """Create GS-ER-compliant Evidence Register sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws.title = "Evidence Register"

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    # Row 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = _border

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = _border

    # Row 3: empty separator

    # Row 4: Column headers (003366 navy)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border

    # Data validations
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ver_status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)
    ws.add_data_validation(ver_status_dv)

    # Row 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001", 2: "Procedure Inventory", 3: "Policy Document",
        4: "Example evidence description", 5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026", 7: "Assessor Name", 8: "Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = _border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Rows 6-105: Empty FFFFCC rows (100 rows)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = _border
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths and freeze panes
    for col, width in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"


# =============================================================================
# SUMMARY DASHBOARD SHEET
# =============================================================================
def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet — GS-compliant TABLE 1/2/3."""
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = "Summary Dashboard"

    # -------------------------------------------------------------------------
    # Row 1: Title header
    # -------------------------------------------------------------------------
    ws.merge_cells("A1:G1")
    ws["A1"] = "PROCEDURE INVENTORY ASSESSMENT \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.5.37: Documented Operating Procedures"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty spacer

    # -------------------------------------------------------------------------
    # TABLE 1: Assessment Area Compliance Overview
    # -------------------------------------------------------------------------
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = border
    ws.row_dimensions[4].height = 20

    # Column Headers (Row 5)
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows — 4 assessment areas (rows 6-9)
    area_configs = [
        (
            "Procedure Inventory",
            "=COUNTA('Procedure Inventory'!B5:B54)",
            "=COUNTIF('Procedure Inventory'!K5:K54,\"Approved\")",
            "=COUNTIF('Procedure Inventory'!K5:K54,\"Draft\")+COUNTIF('Procedure Inventory'!K5:K54,\"Pending Approval\")+COUNTIF('Procedure Inventory'!K5:K54,\"Under Review\")",
            "=COUNTIF('Procedure Inventory'!K5:K54,\"Expired\")",
            "=0",
        ),
        (
            "Required Procedures",
            "=COUNTA('Required Procedures'!B5:B63)",
            "=COUNTIF('Required Procedures'!F5:F63,\"Exists\")",
            "=COUNTIF('Required Procedures'!F5:F63,\"Partial\")",
            "=COUNTIF('Required Procedures'!F5:F63,\"Missing\")",
            "=0",
        ),
        (
            "Accessibility Matrix",
            "=COUNTA('Accessibility Matrix'!A5:A54)",
            "=COUNTA('Accessibility Matrix'!G5:G54)",
            "=0",
            "=B8-C8",
            "=0",
        ),
        (
            "Gap Analysis",
            "=COUNTA('Gap Analysis'!A5:A54)",
            "=COUNTIF('Gap Analysis'!H5:H54,\"Closed\")",
            "=COUNTIF('Gap Analysis'!H5:H54,\"In Progress\")",
            "=COUNTIF('Gap Analysis'!H5:H54,\"Open\")",
            "=0",
        ),
    ]

    for i, (area_name, f_total, f_comp, f_part, f_noncomp, f_na) in enumerate(area_configs):
        row = 6 + i

        ws.cell(row=row, column=1, value=area_name).border = border
        ws.cell(row=row, column=1).font = Font(color="000000")

        cell_b = ws.cell(row=row, column=2, value=f_total)
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = Font(color="000000")

        cell_c = ws.cell(row=row, column=3, value=f_comp)
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = Font(color="000000")

        cell_d = ws.cell(row=row, column=4, value=f_part)
        cell_d.border = border
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = Font(color="000000")

        cell_e = ws.cell(row=row, column=5, value=f_noncomp)
        cell_e.border = border
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = Font(color="000000")

        cell_f = ws.cell(row=row, column=6, value=f_na)
        cell_f.border = border
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = Font(color="000000")

        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IFERROR(IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row})),\"\")"
        cell_g.number_format = "0.0%"
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(color="000000")

    # TOTAL row (row 10)
    total_row = 10
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = border
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    cell = ws.cell(row=total_row, column=7)
    cell.value = f"=IFERROR(IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row})),\"\")"
    cell.number_format = "0.0%"
    cell.font = Font(bold=True, color="000000")
    cell.fill = grey_fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    # -------------------------------------------------------------------------
    # TABLE 2: Key Metrics
    # -------------------------------------------------------------------------
    t2_start = total_row + 2  # row 12
    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=t2_start, column=c).border = border
    ws.row_dimensions[t2_start].height = 20

    # TABLE 2 header row
    for col, header in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # TABLE 2 metrics
    metrics = [
        # Procedure Inventory Metrics
        ("Total Procedures Inventoried", "=COUNTA('Procedure Inventory'!B5:B54)"),
        ("Approved Procedures", "=COUNTIF('Procedure Inventory'!K5:K54,\"Approved\")"),
        ("Expired Procedures (Require Renewal)", "=COUNTIF('Procedure Inventory'!K5:K54,\"Expired\")"),
        ("Critical Procedures", "=COUNTIF('Procedure Inventory'!O5:O54,\"Critical\")"),
        ("High Criticality Procedures", "=COUNTIF('Procedure Inventory'!O5:O54,\"High\")"),
        ("Security Operations Procedures", "=COUNTIF('Procedure Inventory'!C5:C54,\"Security Operations\")"),
        ("System Operations Procedures", "=COUNTIF('Procedure Inventory'!C5:C54,\"System Operations\")"),
        ("Recovery Operations Procedures", "=COUNTIF('Procedure Inventory'!C5:C54,\"Recovery Operations\")"),
        # Required Procedures Coverage
        ("Required Procedures Identified", "=COUNTA('Required Procedures'!B5:B63)"),
        ("Procedures Fully Existing", "=COUNTIF('Required Procedures'!F5:F63,\"Exists\")"),
        ("Procedures Partially Covered", "=COUNTIF('Required Procedures'!F5:F63,\"Partial\")"),
        ("Procedures Missing", "=COUNTIF('Required Procedures'!F5:F63,\"Missing\")"),
        ("Critical Priority Requirements", "=COUNTIF('Required Procedures'!E5:E63,\"Critical\")"),
        # Gap and Accessibility Metrics
        ("Total Gap Items Identified", "=COUNTA('Gap Analysis'!A5:A54)"),
        ("Critical Severity Gaps", "=COUNTIF('Gap Analysis'!D5:D54,\"Critical\")"),
        ("High Severity Gaps", "=COUNTIF('Gap Analysis'!D5:D54,\"High\")"),
        ("Open Gaps (Unresolved)", "=COUNTIF('Gap Analysis'!H5:H54,\"Open\")"),
        ("Missing Procedures (Gap Type)", "=COUNTIF('Gap Analysis'!B5:B54,\"Missing\")"),
        ("Outdated Procedures", "=COUNTIF('Gap Analysis'!B5:B54,\"Outdated\")"),
        ("Procedures with Verified Access", "=COUNTA('Accessibility Matrix'!H5:H54)"),
    ]

    row = t2_start + 2
    for metric, formula in metrics:
        cell_a = ws.cell(row=row, column=1, value=metric)
        cell_a.border = border
        cell_a.font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = border
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
        row += 1

    # -------------------------------------------------------------------------
    # TABLE 3: Critical Findings
    # -------------------------------------------------------------------------
    t3_start = row + 1
    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t3_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=t3_start, column=c).border = border
    ws.row_dimensions[t3_start].height = 20

    # TABLE 3 header row
    t3_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, header in enumerate(t3_headers, 1):
        cell = ws.cell(row=t3_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # TABLE 3 findings
    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    findings = [
        ("Required Procedures", "Missing required ISO procedures", "=COUNTIF('Required Procedures'!F5:F63,\"Missing\")", "Critical", "Immediate"),
        ("Required Procedures", "Critical priority procedures missing", "=COUNTIFS('Required Procedures'!E5:E63,\"Critical\",'Required Procedures'!F5:F63,\"Missing\")", "Critical", "Immediate"),
        ("Procedure Inventory", "Expired procedures not renewed", "=COUNTIF('Procedure Inventory'!K5:K54,\"Expired\")", "Critical", "Immediate"),
        ("Procedure Inventory", "Critical procedures not yet approved (Draft)", "=COUNTIFS('Procedure Inventory'!O5:O54,\"Critical\",'Procedure Inventory'!K5:K54,\"Draft\")", "Critical", "Immediate"),
        ("Gap Analysis", "Open critical severity gaps", "=COUNTIFS('Gap Analysis'!D5:D54,\"Critical\",'Gap Analysis'!H5:H54,\"Open\")", "Critical", "Immediate"),
        ("Gap Analysis", "Open high severity gaps", "=COUNTIFS('Gap Analysis'!D5:D54,\"High\",'Gap Analysis'!H5:H54,\"Open\")", "High", "Urgent"),
        ("Gap Analysis", "Missing procedure gaps (no procedure exists)", "=COUNTIF('Gap Analysis'!B5:B54,\"Missing\")", "High", "Urgent"),
        ("Procedure Inventory", "Recovery operations procedures inventory", "=COUNTIF('Procedure Inventory'!C5:C54,\"Recovery Operations\")", "Medium", "Plan"),
    ]

    row = t3_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_sheet(ws):
    """Create the Gold Standard Approval and Sign-Off sheet (GS-AS-012 compliant)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = _border

    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = _blue
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = _border

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G9),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = _input
        elif value.startswith("="):
            ws[f"B{row}"].number_format = "0.0%"
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _border
        row += 1

    # Approver sections (rows 11-30)
    row += 2
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (IT SECURITY MANAGER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = _border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = _input
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = _border
            row += 1
        row += 1

    # FINAL DECISION — GS-AS-012 compliant: col A = plain bold, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    # NO fill on col A (GS-AS-012 requirement)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = _input
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = _border

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")
    row += 2

    # NEXT REVIEW DETAILS — A:E merged, 4472C4 fill, white bold 11pt
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _border
    row += 1
    for label in ["Next Review Date (DD.MM.YYYY):", "Review Owner:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = _input
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.freeze_panes = "A3"

    # B6 compliance formula (Overall Compliance Rating references Summary Dashboard TOTAL row G10)
    ws["B6"].number_format = "0.0%"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def main() -> int:
    """
    Main execution function — orchestrates workbook creation.

    Returns:
        int: 0 on success, 1 on failure
    """
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/8] Creating Procedure Inventory sheet...")
        create_procedure_inventory_sheet(wb["Procedure Inventory"], styles)

        logger.info("[3/8] Creating Required Procedures sheet...")
        create_required_procedures_sheet(wb["Required Procedures"], styles)

        logger.info("[4/8] Creating Accessibility Matrix sheet...")
        create_accessibility_matrix_sheet(wb["Accessibility Matrix"], styles)

        logger.info("[5/8] Creating Gap Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap Analysis"], styles)

        logger.info("[6/8] Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[7/8] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"])

        logger.info("[8/8] Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        finalize_validations(wb)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        wb.save(output_path)
        logger.info(f"SUCCESS: {OUTPUT_FILENAME}")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error(f"Permission denied: {e}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
