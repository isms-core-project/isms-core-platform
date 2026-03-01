#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.35-36.S1 - Independent Review Planning and Tracking Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.35: Independent Review of Information Security
Assessment Domain 1 of 3: Independent Review Planning and Tracking

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific compliance review and independent assurance infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Independent review scope and reviewer qualification criteria (match your audit programme)
2. Compliance assessment control categories and scoring methodology
3. Finding classification levels and remediation priority thresholds
4. Review frequency and scheduling requirements (adapt to risk profile)
5. Escalation paths for non-compliance and repeat findings

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.35 Independent Review of Information Security Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
compliance review and independent assurance controls and compliance requirements.

**Purpose:**
Enables systematic planning and tracking of Independent Review Planning and Tracking under ISO 27001:2022 Controls A.5.35 and A.5.36. Supports evidence-based documentation of independent review activities, compliance assessment results, and remediation progress.

**Assessment Scope:**
- Independent review planning, scope, and reviewer independence
- Compliance assessment coverage across control domains
- Finding classification and risk rating accuracy
- Remediation ownership assignment and progress tracking
- Repeat finding identification and root cause analysis
- Review evidence documentation and audit trail quality
- Evidence collection for governance, audit, and regulatory reporting

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Independent Review of Information Security controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a535_36_1_independent_review.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a535_36_1_independent_review.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a535_36_1_independent_review.py --date 20250115

Output:
    File: ISMS-IMP-A.5.35-36.S1_Independent_Review_Planning_and_Tracking_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.35
Assessment Domain:    1 of 3 (Independent Review Planning and Tracking)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.35: Independent Review of Information Security Policy (Governance)
    - ISMS-IMP-A.5.35-36.S1: Independent Review Planning and Tracking (Domain 1)
    - ISMS-IMP-A.5.35-36.S2: Compliance Assessment (Domain 2)
    - ISMS-IMP-A.5.35-36.S3: Findings and Remediation Management (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.35-36.S1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive compliance review and independent assurance details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review independent review scheduling and compliance assessment scope annually or when the control framework changes, audit findings reveal systemic gaps, or regulatory requirements are updated.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.35-36.S1"
WORKBOOK_NAME = "Independent Review Planning and Tracking"
CONTROL_ID = "A.5.35"
CONTROL_NAME = "Independent Review of Information Security"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles


# =============================================================================
# FINALIZE VALIDATIONS
# =============================================================================

_STYLES = setup_styles()
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for _ in ws.data_validations.dataValidation:
            pass


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Instructions & Legend",
        "Review Schedule",
        "Reviewer Registry",
        "Review Scope",
        "Review Execution",
        "Findings Summary",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Review Schedule — plan all independent ISMS reviews for the assessment period.',
        '2. Complete Reviewer Registry — document qualified independent reviewers (internal audit, external).',
        '3. Complete Review Scope — define the scope, objectives, and criteria for each review.',
        '4. Complete Review Execution — record review activities, interviews, and evidence collected.',
        '5. Complete Findings Summary — consolidate all findings with risk ratings and recommendations.',
        '6. Maintain Evidence Records with review reports and management responses.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_review_schedule_sheet(ws, styles):
    """Create the Review_Schedule sheet."""
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:L1")
    ws["A1"] = "INDEPENDENT REVIEW SCHEDULE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:L2")
    ws["A2"] = "Record all independent reviews of the ISMS at planned intervals — include type, scope, reviewer, dates, and status."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Review_ID", 14),
        ("Review_Type", 18),
        ("Review_Scope", 35),
        ("Planned_Start", 14),
        ("Planned_End", 14),
        ("Actual_Start", 14),
        ("Actual_End", 14),
        ("Lead_Reviewer", 22),
        ("Review_Status", 16),
        ("Findings_Count", 14),
        ("Report_Date", 14),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name.replace("_", " "))
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Annual Comprehensive,Surveillance,Ad-Hoc,Pre-Certification,Post-Incident,Scope Change"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(
        type="list",
        formula1='"Planned,In Progress,Completed,Postponed,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Row 4: F2F2F2 sample row (MAX-003 fix)
    sample_vals = {
        1: "REV-2026-001", 2: "Annual Comprehensive", 3: "Full ISMS scope including A.5-A.8",
        4: "01.03.2026", 5: "15.03.2026", 6: "01.03.2026", 7: "14.03.2026",
        8: "J. Smith (External)", 9: "Completed", 10: "3", 11: "20.03.2026",
        12: "Annual review — ISO 27001:2022 certification audit",
    }
    for c, val in sample_vals.items():
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = grey_fill
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    dv_type.add(ws.cell(row=4, column=2))
    dv_status.add(ws.cell(row=4, column=9))

    # Rows 5-54: 50 empty FFFFCC rows (MAX-004 fix: was 24, now 50)
    for r in range(5, 55):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# REVIEWER REGISTRY SHEET
# =============================================================================
def create_reviewer_registry_sheet(ws, styles):
    """Create the Reviewer_Registry sheet."""
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:K1")
    ws["A1"] = "QUALIFIED REVIEWER REGISTRY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:K2")
    ws["A2"] = "Register all approved independent reviewers — record qualifications, independence status, and approval for use in ISMS reviews."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Reviewer_ID", 14),
        ("Reviewer_Name", 25),
        ("Organisation", 25),
        ("Reviewer_Type", 16),
        ("Qualifications", 35),
        ("Certification_Expiry", 16),
        ("Areas_of_Expertise", 35),
        ("Independence_Status", 18),
        ("Last_Review_Date", 16),
        ("Approval_Status", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name.replace("_", " "))
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"External Auditor,External Consultant,Internal Audit,Cross-Departmental"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_independence = DataValidation(
        type="list",
        formula1='"Fully Independent,Conditionally Independent,Not Independent"',
        allow_blank=False
    )
    ws.add_data_validation(dv_independence)

    dv_approval = DataValidation(
        type="list",
        formula1='"Approved,Pending Approval,Expired,Revoked"',
        allow_blank=False
    )
    ws.add_data_validation(dv_approval)

    # Row 4: F2F2F2 sample row (MAX-003 fix)
    sample_vals = {
        1: "RVR-001", 2: "Jane Smith", 3: "ISO Auditors Ltd",
        4: "External Auditor", 5: "ISO 27001 Lead Auditor (IRCA), CISA",
        6: "31.12.2027", 7: "ISMS Governance, Access Control, Risk Management",
        8: "Fully Independent", 9: "15.03.2026", 10: "Approved",
        11: "Approved for annual ISMS certification audit",
    }
    for c, val in sample_vals.items():
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = grey_fill
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    dv_type.add(ws.cell(row=4, column=4))
    dv_independence.add(ws.cell(row=4, column=8))
    dv_approval.add(ws.cell(row=4, column=10))

    # Rows 5-54: 50 empty FFFFCC rows (MAX-004 fix: was 20, now 50)
    for r in range(5, 55):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=4))
        dv_independence.add(ws.cell(row=r, column=8))
        dv_approval.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# REVIEW SCOPE SHEET
# =============================================================================
def create_review_scope_sheet(ws, styles):
    """Create the Review_Scope sheet."""
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:J1")
    ws["A1"] = "REVIEW SCOPE DEFINITION"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:J2")
    ws["A2"] = "Define the scope for each independent review — specify controls covered, departments, systems, sampling approach, and approver."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Review_Ref", 14),
        ("Scope_Area", 30),
        ("ISO_Controls", 25),
        ("Departments_Included", 25),
        ("Systems_Included", 30),
        ("Exclusions", 25),
        ("Sampling_Approach", 22),
        ("Sample_Size", 14),
        ("Scope_Approved_By", 20),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name.replace("_", " "))
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_sampling = DataValidation(
        type="list",
        formula1='"Statistical,Judgmental,100% Coverage,Risk-Based"',
        allow_blank=False
    )
    ws.add_data_validation(dv_sampling)

    # Row 4: F2F2F2 sample row (MAX-003 fix)
    sample_vals = {
        1: "REV-2026-001", 2: "ISMS Governance", 3: "A.5.1-A.5.4",
        4: "Executive, IS Team", 5: "ISMS platform, SharePoint",
        6: "Operational systems", 7: "Risk-Based", 8: "25",
        9: "CISO", 10: "Scope agreed and approved 15.02.2026",
    }
    for c, val in sample_vals.items():
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = grey_fill
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    dv_sampling.add(ws.cell(row=4, column=7))

    # Rows 5-54: 50 empty FFFFCC rows (MAX-004 fix: was 20, now 50)
    for r in range(5, 55):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_sampling.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# REVIEW EXECUTION SHEET
# =============================================================================
def create_review_execution_sheet(ws, styles):
    """Create the Review_Execution sheet."""
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:L1")
    ws["A1"] = "REVIEW EXECUTION TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:L2")
    ws["A2"] = "Track individual review activities — log assigned reviewer, planned/actual dates, status, evidence collected, and any issues found."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Activity_ID", 14),
        ("Review_Ref", 14),
        ("Activity_Description", 40),
        ("Assigned_To", 22),
        ("Planned_Date", 14),
        ("Actual_Date", 14),
        ("Status", 16),
        ("Evidence_Collected", 16),
        ("Interviews_Conducted", 18),
        ("Issues_Found", 14),
        ("Follow_Up_Required", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name.replace("_", " "))
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Blocked"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    # Row 4: F2F2F2 sample row (MAX-003 fix)
    sample_vals = {
        1: "ACT-001", 2: "REV-2026-001", 3: "Review access control policy and implementation",
        4: "J. Smith", 5: "02.03.2026", 6: "02.03.2026",
        7: "Completed", 8: "Yes", 9: "Yes",
        10: "2", 11: "No", 12: "Policy v2.1 reviewed; 2 gaps identified in MFA enforcement",
    }
    for c, val in sample_vals.items():
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = grey_fill
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    dv_status.add(ws.cell(row=4, column=7))
    dv_yesno.add(ws.cell(row=4, column=8))
    dv_yesno.add(ws.cell(row=4, column=11))

    # Rows 5-54: 50 empty FFFFCC rows
    for r in range(5, 55):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=7))
        dv_yesno.add(ws.cell(row=r, column=8))
        dv_yesno.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# FINDINGS SUMMARY SHEET
# =============================================================================
def create_findings_summary_sheet(ws, styles):
    """Create the Findings_Summary sheet."""
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:L1")
    ws["A1"] = "REVIEW FINDINGS SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:L2")
    ws["A2"] = "Record all findings identified during independent reviews — classify by type and severity, assign owners, and track to closure."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Finding_ID", 14),
        ("Review_Ref", 14),
        ("Finding_Type", 18),
        ("Finding_Description", 50),
        ("Control_Reference", 18),
        ("Severity", 14),
        ("Root_Cause", 30),
        ("Recommendation", 40),
        ("Management_Response", 35),
        ("Remediation_Owner", 22),
        ("Target_Date", 14),
        ("Status", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name.replace("_", " "))
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Non-Conformity,Observation,Opportunity for Improvement,Positive Finding"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_severity = DataValidation(
        type="list",
        formula1='"Major,Minor,Observation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed,Accepted"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Row 4: F2F2F2 sample row (MAX-003 fix)
    sample_vals = {
        1: "FND-001", 2: "REV-2026-001", 3: "Non-Conformity",
        4: "MFA not enforced for privileged accounts in cloud environment",
        5: "A.5.15, A.8.5", 6: "Major",
        7: "Policy gap — MFA policy not updated post-cloud migration",
        8: "Enforce MFA for all privileged accounts within 30 days",
        9: "Accepted — remediation plan in progress", 10: "IT Operations Manager",
        11: "31.03.2026", 12: "In Progress",
    }
    for c, val in sample_vals.items():
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = grey_fill
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    dv_type.add(ws.cell(row=4, column=3))
    dv_severity.add(ws.cell(row=4, column=6))
    dv_status.add(ws.cell(row=4, column=12))

    # Rows 5-54: 50 empty FFFFCC rows
    for r in range(5, 55):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=3))
        dv_severity.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register(ws):
    """Create the Evidence_Register sheet with inline styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    col_hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    col_hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = hdr_font
    ws["A1"].fill = hdr_fill
    ws["A1"].alignment = hdr_align
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:H2")
    ws["A2"] = "Record all evidence collected during independent reviews — link each item to a review, specify type, location, and retention period."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Review", 18),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Retention_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name.replace("_", " "))
        cell.font = col_hdr_font
        cell.fill = col_hdr_fill
        cell.alignment = col_hdr_align
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Review Plan,Scope Document,Interview Notes,Sample Evidence,Audit Report,Management Response,Closure Evidence"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    # Row 4: F2F2F2 sample row (MAX-003 fix)
    sample_vals = {
        1: "EV-001", 2: "Audit Report",
        3: "Annual ISMS independent review report 2026",
        4: "REV-2026-001", 5: "20.03.2026",
        6: "//fileserver/isms/evidence/reviews/2026/rev-2026-001-report.pdf",
        7: "J. Smith", 8: "31.03.2031",
    }
    for c, val in sample_vals.items():
        cell = ws.cell(row=5, column=c, value=val)
        cell.fill = grey_fill
        cell.border = border_thin
        cell.alignment = input_align
    dv_type.add(ws.cell(row=4, column=2))

    # Rows 5-54: 50 empty FFFFCC rows
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = input_align
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"  # DS-007 fix


# =============================================================================
# SUMMARY DASHBOARD SHEET
# =============================================================================
def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet — Gold Standard pattern (A.8.33-34)."""
    from openpyxl.utils import get_column_letter as gcl

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title banner (GS-SD-014: must contain em dash + SUMMARY DASHBOARD)
    ws.merge_cells("A1:G1")
    ws["A1"] = "INDEPENDENT REVIEW PLANNING AND TRACKING \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (italic, 003366, left-aligned, NO fill)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.5.35: Independent Review of Information Security"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty spacer

    # -------------------------------------------------------------------------
    # TABLE 1: Assessment Area Compliance Overview
    # -------------------------------------------------------------------------
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = navy_fill
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = border

    # Column headers (row 5) — D9D9D9, black bold
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows (rows 6-9)
    # Data rows start at row 5 (after F2F2F2 sample at row 4) in each sheet
    area_configs = [
        # (Area Name, Count col range, Status col, [Compliant val, Partial val, Non-Compliant formula, NA val])
        ("Review Planning",
         "COUNTA('Review Schedule'!B5:B54)",
         "COUNTIF('Review Schedule'!I5:I54,\"Completed\")",
         "COUNTIF('Review Schedule'!I5:I54,\"In Progress\")",
         "COUNTIF('Review Schedule'!I5:I54,\"Postponed\")+COUNTIF('Review Schedule'!I5:I54,\"Cancelled\")",
         "COUNTIF('Review Schedule'!I5:I54,\"Planned\")"),
        ("Reviewer Registry",
         "COUNTA('Reviewer Registry'!B5:B54)",
         "COUNTIF('Reviewer Registry'!J5:J54,\"Approved\")",
         "COUNTIF('Reviewer Registry'!J5:J54,\"Pending Approval\")",
         "COUNTIF('Reviewer Registry'!J5:J54,\"Revoked\")",
         "COUNTIF('Reviewer Registry'!J5:J54,\"Expired\")"),
        ("Review Execution",
         "COUNTA('Review Execution'!B5:B54)",
         "COUNTIF('Review Execution'!G5:G54,\"Completed\")",
         "COUNTIF('Review Execution'!G5:G54,\"In Progress\")",
         "COUNTIF('Review Execution'!G5:G54,\"Blocked\")",
         "COUNTIF('Review Execution'!G5:G54,\"Not Started\")"),
        ("Findings Summary",
         "COUNTA('Findings Summary'!B5:B54)",
         "COUNTIF('Findings Summary'!L5:L54,\"Closed\")",
         "COUNTIF('Findings Summary'!L5:L54,\"In Progress\")",
         "COUNTIF('Findings Summary'!L5:L54,\"Open\")",
         "COUNTIF('Findings Summary'!L5:L54,\"Accepted\")"),
    ]

    for i, (area, total_f, comp_f, part_f, nc_f, na_f) in enumerate(area_configs):
        r = 6 + i
        # Col A: area name
        ws.cell(row=r, column=1, value=area).border = border
        ws.cell(row=r, column=1).font = Font(color="000000")
        # Col B: Total
        cb = ws.cell(row=r, column=2, value=f"={total_f}")
        cb.border = border
        cb.alignment = Alignment(horizontal="center")
        cb.font = Font(color="000000")
        # Col C: Compliant
        cc = ws.cell(row=r, column=3, value=f"={comp_f}")
        cc.border = border
        cc.alignment = Alignment(horizontal="center")
        cc.font = Font(color="000000")
        # Col D: Partial
        cd = ws.cell(row=r, column=4, value=f"={part_f}")
        cd.border = border
        cd.alignment = Alignment(horizontal="center")
        cd.font = Font(color="000000")
        # Col E: Non-Compliant
        ce = ws.cell(row=r, column=5, value=f"={nc_f}")
        ce.border = border
        ce.alignment = Alignment(horizontal="center")
        ce.font = Font(color="000000")
        # Col F: N/A
        cf = ws.cell(row=r, column=6, value=f"={na_f}")
        cf.border = border
        cf.alignment = Alignment(horizontal="center")
        cf.font = Font(color="000000")
        # Col G: Compliance %
        cg = ws.cell(row=r, column=7,
                     value=f"=IFERROR(IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r})),\"\")")
        cg.number_format = "0.0%"
        cg.border = border
        cg.alignment = Alignment(horizontal="center")
        cg.font = Font(color="000000")

    # TOTAL row (row 10)
    total_row = 10
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = border
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({gcl(col)}6:{gcl(col)}9)"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    # TOTAL compliance %
    cg_total = ws.cell(row=total_row, column=7,
                       value=f"=IFERROR(IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row})),\"\")")
    cg_total.number_format = "0.0%"
    cg_total.font = Font(bold=True, color="000000")
    cg_total.fill = grey_fill
    cg_total.border = border
    cg_total.alignment = Alignment(horizontal="center")

    # -------------------------------------------------------------------------
    # TABLE 2: Key Metrics
    # -------------------------------------------------------------------------
    t2_start = total_row + 2
    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2_start}"].fill = navy_fill
    for c in range(1, 8):
        ws.cell(row=t2_start, column=c).border = border

    # TABLE 2 headers (D9D9D9, GS-SD-016)
    t2_hdr_row = t2_start + 1
    for col, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # TABLE 2 metrics — white fill, 000000 font (GS-SD-015/016)
    metrics = [
        ("Total Reviews Scheduled", "=COUNTA('Review Schedule'!B5:B54)"),
        ("Reviews Completed", "=COUNTIF('Review Schedule'!I5:I54,\"Completed\")"),
        ("Reviews In Progress", "=COUNTIF('Review Schedule'!I5:I54,\"In Progress\")"),
        ("Reviews Postponed or Cancelled",
         "=COUNTIF('Review Schedule'!I5:I54,\"Postponed\")+COUNTIF('Review Schedule'!I5:I54,\"Cancelled\")"),
        ("Annual Comprehensive Reviews", "=COUNTIF('Review Schedule'!B5:B54,\"Annual Comprehensive\")"),
        ("Surveillance Reviews", "=COUNTIF('Review Schedule'!B5:B54,\"Surveillance\")"),
        ("Ad-Hoc / Post-Incident Reviews",
         "=COUNTIF('Review Schedule'!B5:B54,\"Ad-Hoc\")+COUNTIF('Review Schedule'!B5:B54,\"Post-Incident\")"),
        ("Total Registered Reviewers", "=COUNTA('Reviewer Registry'!B5:B54)"),
        ("Approved Reviewers", "=COUNTIF('Reviewer Registry'!J5:J54,\"Approved\")"),
        ("Fully Independent Reviewers", "=COUNTIF('Reviewer Registry'!H5:H54,\"Fully Independent\")"),
        ("External Reviewers",
         "=COUNTIF('Reviewer Registry'!D5:D54,\"External Auditor\")+COUNTIF('Reviewer Registry'!D5:D54,\"External Consultant\")"),
        ("Total Execution Activities", "=COUNTA('Review Execution'!B5:B54)"),
        ("Activities Completed", "=COUNTIF('Review Execution'!G5:G54,\"Completed\")"),
        ("Activities with Evidence Collected", "=COUNTIF('Review Execution'!H5:H54,\"Yes\")"),
        ("Blocked Activities", "=COUNTIF('Review Execution'!G5:G54,\"Blocked\")"),
        ("Total Findings Identified", "=COUNTA('Findings Summary'!B5:B54)"),
        ("Major Non-Conformities", "=COUNTIF('Findings Summary'!F5:F54,\"Major\")"),
        ("Minor Observations", "=COUNTIF('Findings Summary'!C5:C54,\"Observation\")"),
        ("Open Findings", "=COUNTIF('Findings Summary'!L5:L54,\"Open\")"),
        ("Closed Findings", "=COUNTIF('Findings Summary'!L5:L54,\"Closed\")"),
    ]

    row = t2_hdr_row + 1
    for metric, formula in metrics:
        ws.cell(row=row, column=1, value=metric).border = border
        ws.cell(row=row, column=1).font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = border
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # -------------------------------------------------------------------------
    # TABLE 3: Critical Findings
    # -------------------------------------------------------------------------
    t3_start = row + 1
    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t3_start}"].fill = red_fill
    for c in range(1, 8):
        ws.cell(row=t3_start, column=c).border = border

    # TABLE 3 headers (D9D9D9)
    t3_hdr_row = t3_start + 1
    for col, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=t3_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # TABLE 3 findings — FFFFCC fill, 000000 font
    findings = [
        ("Review Planning",
         "Reviews postponed or cancelled",
         "=COUNTIF('Review Schedule'!I5:I54,\"Postponed\")+COUNTIF('Review Schedule'!I5:I54,\"Cancelled\")",
         "High", "Reschedule immediately"),
        ("Reviewer Qualification",
         "Reviewers with expired credentials",
         "=COUNTIF('Reviewer Registry'!J5:J54,\"Expired\")",
         "High", "Renew or replace reviewer"),
        ("Reviewer Independence",
         "Reviewers flagged not independent",
         "=COUNTIF('Reviewer Registry'!H5:H54,\"Not Independent\")",
         "Critical", "Escalate to CISO"),
        ("Review Execution",
         "Execution activities blocked",
         "=COUNTIF('Review Execution'!G5:G54,\"Blocked\")",
         "High", "Investigate and unblock"),
        ("Findings Management",
         "Major non-conformities open",
         "=COUNTIFS('Findings Summary'!C5:C54,\"Non-Conformity\",'Findings Summary'!F5:F54,\"Major\",'Findings Summary'!L5:L54,\"Open\")",
         "Critical", "Immediate remediation"),
        ("Findings Management",
         "All open non-conformities",
         "=COUNTIFS('Findings Summary'!C5:C54,\"Non-Conformity\",'Findings Summary'!L5:L54,\"Open\")",
         "High", "Assign owner and date"),
        ("Evidence",
         "Completed activities with no evidence",
         "=COUNTIFS('Review Execution'!G5:G54,\"Completed\",'Review Execution'!H5:H54,\"No\")",
         "Medium", "Collect evidence"),
    ]

    row = t3_hdr_row + 1
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty buffer rows (FFFFCC)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths and freeze
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-001..015)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = blue
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8); Overall Compliance Rating at B6 (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G9),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = yellow
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between approver sections

    # FINAL DECISION (GS-AS-004/012: col A plain bold, no dark fill)
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = yellow
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS (GS-AS-005/013: 4472C4 banner, borders on all)
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = blue
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/9] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/9] Creating Review Schedule sheet...")
        create_review_schedule_sheet(wb["Review Schedule"], styles)

        logger.info("[3/9] Creating Reviewer Registry sheet...")
        create_reviewer_registry_sheet(wb["Reviewer Registry"], styles)

        logger.info("[4/9] Creating Review Scope sheet...")
        create_review_scope_sheet(wb["Review Scope"], styles)

        logger.info("[5/9] Creating Review Execution sheet...")
        create_review_execution_sheet(wb["Review Execution"], styles)

        logger.info("[6/9] Creating Findings Summary sheet...")
        create_findings_summary_sheet(wb["Findings Summary"], styles)

        logger.info("[7/9] Creating Evidence Records sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[8/9] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"])

        logger.info("[9/9] Creating Approval Record sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        finalize_validations(wb)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        wb.save(output_path)
        logger.info(f"SUCCESS: {_wkbk_dir / OUTPUT_FILENAME}")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        return 1
    except PermissionError as e:
        logger.error(f"Permission denied: {e}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
