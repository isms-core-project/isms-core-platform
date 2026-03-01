#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.10-11.S2 - Usage Rules Inventory Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.10: Acceptable Use of Information and Other Associated Assets
Assessment Domain 2 of 3: Usage Rules Inventory

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific asset usage and lifecycle management infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Asset categories and acceptable use rule definitions (match your asset types)
2. Usage monitoring scope and exception categories
3. Asset return and offboarding workflow stages
4. Policy acknowledgment and training tracking mechanisms
5. Disposal and sanitisation procedure references

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.10 Acceptable Use of Information and Other Associated Assets Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
asset usage and lifecycle management controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Usage Rules Inventory compliance under ISO 27001:2022 Controls A.5.10 and A.5.11. Supports evidence-based documentation of acceptable use adherence, asset lifecycle tracking, and return procedures for audit readiness.

**Assessment Scope:**
- Acceptable use policy coverage across asset categories
- Usage rule documentation completeness and currency
- Policy acknowledgment rates and training compliance
- Asset return process adherence during offboarding
- Monitoring and enforcement mechanism effectiveness
- Exception handling and escalation documentation
- Evidence collection for HR and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Acceptable Use of Information and Other Associated Assets controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a510_11_2_usage_rules_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a510_11_2_usage_rules_inventory.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a510_11_2_usage_rules_inventory.py --date 20250115

Output:
    File: ISMS-IMP-A.5.10-11.S2_Usage_Rules_Inventory_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.10
Assessment Domain:    2 of 3 (Usage Rules Inventory)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.10: Acceptable Use of Information and Other Associated Assets Policy (Governance)
    - ISMS-IMP-A.5.10-11.S1: Acceptable Use Policy Assessment (Domain 1)
    - ISMS-IMP-A.5.10-11.S2: Usage Rules Inventory (Domain 2)
    - ISMS-IMP-A.5.10-11.S3: Asset Return and Offboarding Assessment (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.10-11.S2 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive asset usage and lifecycle management details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review acceptable use rules and asset return procedures annually or when asset categories change, new usage patterns emerge, or regulatory requirements are updated.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.10-11.S2"
WORKBOOK_NAME = "Usage Rules Inventory"
CONTROL_ID = "A.5.10"
CONTROL_NAME = "Acceptable Use of Information and Other Associated Assets"
CONTROL_REF = f"ISO/IEC 27001:2022 \u2014 Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": THIN_BORDER,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": THIN_BORDER,
        },
        "border": THIN_BORDER,
        "permitted": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "prohibited": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "conditional": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
    }
    return styles


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

_STYLES = setup_styles()
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Instructions & Legend",
        "Usage Rules",
        "Permitted Activities",
        "Prohibited Activities",
        "Handling Requirements",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Usage Rules — document specific rules for each asset type (devices, systems, data).',
        '2. Complete Permitted Activities — list explicitly permitted uses per asset category.',
        '3. Complete Prohibited Activities — list explicitly prohibited uses with consequence framework.',
        '4. Complete Handling Requirements — document classification-based handling procedures.',
        '5. Maintain Evidence Records with policy documentation and acknowledgement records.',
        '6. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A20"] = "Status Legend"
    ws["A20"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=21, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 22 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_usage_rules_sheet(ws, styles):
    """Create the Usage Rules sheet - master inventory."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "USAGE RULES INVENTORY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:L2")
    ws["A2"] = "Document acceptable use rules for each asset category per ISO 27001:2022 A.5.10"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Rule_ID", 14),
        ("Asset_Category", 25),
        ("Rule_Description", 50),
        ("Classification", 25),
        ("Applies_To", 22),
        ("Enforcement_Method", 22),
        ("Monitoring_Required", 18),
        ("Exception_Process", 18),
        ("Policy_Reference", 20),
        ("Last_Updated", 14),
        ("Owner", 22),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_classification = DataValidation(
        type="list",
        formula1='"Permitted,Permitted with Conditions,Prohibited,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(dv_classification)

    dv_monitoring = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_monitoring)

    # Row 4: F2F2F2 sample row
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = [
        "UR-001",
        "Email",
        "Business email for work communications only; personal use limited to 10 min/day",
        "Permitted with Conditions",
        "All Employees",
        "Email gateway monitoring",
        "Yes",
        "Formal exception request to IT Security",
        "AUP Section 4.1",
        "2026-01-15",
        "IT Security Manager",
        "Reviewed annually",
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = sample_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Pre-populate common usage rules (rows 5-9)
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    usage_rules = [
        ("UR-001", "Email", "Business email for work communications only", "Permitted with Conditions"),
        ("UR-002", "Email", "No forwarding of confidential data to personal accounts", "Prohibited"),
        ("UR-003", "Internet", "Web browsing for business research permitted", "Permitted"),
        ("UR-004", "Internet", "No access to prohibited website categories", "Prohibited"),
        ("UR-005", "Software", "Only approved software may be installed", "Permitted with Conditions"),
    ]

    for row_idx, (rule_id, category, description, classification) in enumerate(usage_rules, start=5):
        ws.cell(row=row_idx, column=1, value=rule_id).fill = grey_fill
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=category).fill = grey_fill
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=description).fill = grey_fill
        ws.cell(row=row_idx, column=3).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=classification).fill = grey_fill
        ws.cell(row=row_idx, column=4).border = THIN_BORDER

        for c in range(5, 13):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = THIN_BORDER
            cell.alignment = styles["input_cell"]["alignment"]

        dv_classification.add(ws.cell(row=row_idx, column=4))
        dv_monitoring.add(ws.cell(row=row_idx, column=7))

    # Empty FFFFCC rows (50 empty rows after sample + pre-pop: rows 10-59)
    for r in range(len(usage_rules) + 5, 60):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = THIN_BORDER
            cell.alignment = styles["input_cell"]["alignment"]
        dv_classification.add(ws.cell(row=r, column=4))
        dv_monitoring.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"


# =============================================================================
# PERMITTED ACTIVITIES SHEET
# =============================================================================
def create_permitted_activities_sheet(ws, styles):
    """Create the Permitted Activities sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "PERMITTED ACTIVITIES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all permitted activities with conditions and approval requirements"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Activity_ID", 14),
        ("Asset_Category", 22),
        ("Permitted_Activity", 45),
        ("Conditions", 35),
        ("Approval_Required", 16),
        ("Approver_Role", 22),
        ("Time_Restrictions", 20),
        ("Location_Restrictions", 22),
        ("Documentation_Required", 20),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_approval = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_approval)

    # Row 4: F2F2F2 sample row
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = [
        "PA-001",
        "Internet",
        "Access approved cloud SaaS tools for business purposes",
        "Must use corporate account; no personal cloud storage",
        "No",
        "N/A",
        "Business hours only",
        "Office and approved remote locations",
        "No",
        "Covered by AUP Section 4.2",
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = sample_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Empty FFFFCC rows (rows 5-54 = 50 rows)
    for r in range(5, 55):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = THIN_BORDER
            cell.alignment = styles["input_cell"]["alignment"]
        dv_approval.add(ws.cell(row=r, column=5))
        dv_approval.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"


# =============================================================================
# PROHIBITED ACTIVITIES SHEET
# =============================================================================
def create_prohibited_activities_sheet(ws, styles):
    """Create the Prohibited Activities sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "PROHIBITED ACTIVITIES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all prohibited activities with risk levels and detection methods"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Prohibition_ID", 14),
        ("Asset_Category", 22),
        ("Prohibited_Activity", 45),
        ("Reason", 35),
        ("Risk_Level", 14),
        ("Detection_Method", 25),
        ("Consequence", 25),
        ("Exception_Possible", 18),
        ("Related_Control", 18),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_risk)

    dv_exception = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_exception)

    # Row 4: F2F2F2 sample row
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = [
        "PH-001",
        "All Assets",
        "Sharing credentials with others (passwords, tokens, certificates)",
        "Account compromise and audit trail loss",
        "Critical",
        "SIEM alerts, login anomaly detection",
        "Formal warning; potential dismissal",
        "No",
        "A.5.17",
        "Zero tolerance policy",
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = sample_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Pre-populate common prohibitions (rows 5-9)
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    prohibitions = [
        ("PH-001", "All Assets", "Sharing credentials with others", "Security breach risk", "Critical"),
        ("PH-002", "All Assets", "Disabling security controls", "Compliance violation", "Critical"),
        ("PH-003", "Email", "Sending unencrypted sensitive data externally", "Data breach risk", "High"),
        ("PH-004", "Software", "Installing unauthorised software", "Malware risk", "High"),
        ("PH-005", "Internet", "Accessing illegal or inappropriate content", "Legal/HR risk", "High"),
    ]

    for row_idx, (p_id, category, activity, reason, risk) in enumerate(prohibitions, start=5):
        ws.cell(row=row_idx, column=1, value=p_id).fill = grey_fill
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=category).fill = grey_fill
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=activity).fill = grey_fill
        ws.cell(row=row_idx, column=3).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=reason).fill = grey_fill
        ws.cell(row=row_idx, column=4).border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=risk).fill = grey_fill
        ws.cell(row=row_idx, column=5).border = THIN_BORDER

        for c in range(6, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = THIN_BORDER
            cell.alignment = styles["input_cell"]["alignment"]

        dv_risk.add(ws.cell(row=row_idx, column=5))
        dv_exception.add(ws.cell(row=row_idx, column=8))

    # Empty FFFFCC rows (50 empty rows after sample + pre-pop: rows 10-59)
    for r in range(len(prohibitions) + 5, 60):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = THIN_BORDER
            cell.alignment = styles["input_cell"]["alignment"]
        dv_risk.add(ws.cell(row=r, column=5))
        dv_exception.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


# =============================================================================
# HANDLING REQUIREMENTS SHEET
# =============================================================================
def create_handling_requirements_sheet(ws, styles):
    """Create the Handling Requirements sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "ASSET HANDLING REQUIREMENTS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:K2")
    ws["A2"] = "Define handling requirements for each asset category by data classification"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Handling_ID", 14),
        ("Asset_Category", 22),
        ("Data_Classification", 18),
        ("Storage_Requirement", 30),
        ("Transmission_Requirement", 30),
        ("Disposal_Requirement", 30),
        ("Access_Restriction", 25),
        ("Labelling_Required", 16),
        ("Encryption_Required", 16),
        ("Retention_Period", 18),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_classification = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=False
    )
    ws.add_data_validation(dv_classification)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    # Row 4: F2F2F2 sample row
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = [
        "HR-001",
        "Information Assets",
        "Restricted",
        "Encrypted storage on approved systems only",
        "TLS 1.2+ encryption in transit; no email",
        "Secure deletion (DoD 5220.22-M or equivalent)",
        "Need-to-know basis; CISO approval",
        "Yes",
        "Yes",
        "7 years (regulatory)",
        "PII data — GDPR/nFADP applies",
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = sample_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Empty FFFFCC rows (rows 5-54 = 50 rows)
    for r in range(5, 55):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = THIN_BORDER
            cell.alignment = styles["input_cell"]["alignment"]
        dv_classification.add(ws.cell(row=r, column=3))
        dv_yesno.add(ws.cell(row=r, column=8))
        dv_yesno.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"


# =============================================================================
# SUMMARY DASHBOARD SHEET
# =============================================================================
def create_summary_dashboard_sheet(ws):
    """Create Gold Standard Summary Dashboard for ISMS-IMP-A.5.10-11.S2."""
    navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title banner (GS-SD-014: em dash + SUMMARY DASHBOARD)
    ws.merge_cells("A1:G1")
    ws["A1"] = "USAGE RULES INVENTORY \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = THIN_BORDER
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (horizontal="left", NO fill)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.5.10: Acceptable Use of Information and Other Associated Assets"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty spacer

    # TABLE 1 banner (Row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = navy_fill
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 20

    # TABLE 1 column headers (Row 5)
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
                  "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows — S2 is documentation focused: totals + prohibited items
    # Usage Rules: sample row 4, 5 pre-pop rows 5-9, 50 empty FFFFCC rows 10-59
    # Prohibited Activities: sample row 4, 5 pre-pop rows 5-9, 50 empty FFFFCC rows 10-59
    # Permitted Activities: sample row 4, 50 empty FFFFCC rows 5-54 (no pre-pop)
    # Handling Requirements: sample row 4, 50 empty FFFFCC rows 5-54 (no pre-pop)
    area_configs = [
        # (Area Name, Total, Compliant, Partial, NonCompliant, NA)
        (
            "Usage Rules Documented",
            "=COUNTA('Usage Rules'!B5:B59)",
            "=COUNTIF('Usage Rules'!D5:D59,\"Permitted\")+COUNTIF('Usage Rules'!D5:D59,\"Permitted with Conditions\")",
            "=0",
            "=COUNTIF('Usage Rules'!D5:D59,\"Not Applicable\")",
            "=0",
        ),
        (
            "Prohibited Activities Defined",
            "=COUNTA('Prohibited Activities'!A5:A59)",
            "=COUNTA('Prohibited Activities'!A5:A59)",
            "=0",
            "=0",
            "=0",
        ),
        (
            "Handling Requirements Defined",
            "=COUNTA('Handling Requirements'!A5:A54)",
            "=COUNTA('Handling Requirements'!A5:A54)",
            "=0",
            "=0",
            "=0",
        ),
    ]

    for i, (area_name, total_f, comp_f, part_f, ncomp_f, na_f) in enumerate(area_configs):
        row = 6 + i

        cell_a = ws.cell(row=row, column=1, value=area_name)
        cell_a.border = THIN_BORDER
        cell_a.font = Font(color="000000")

        cell_b = ws.cell(row=row, column=2, value=total_f)
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = Font(color="000000")

        cell_c = ws.cell(row=row, column=3, value=comp_f)
        cell_c.border = THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = Font(color="000000")

        cell_d = ws.cell(row=row, column=4, value=part_f)
        cell_d.border = THIN_BORDER
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = Font(color="000000")

        cell_e = ws.cell(row=row, column=5, value=ncomp_f)
        cell_e.border = THIN_BORDER
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = Font(color="000000")

        cell_f = ws.cell(row=row, column=6, value=na_f)
        cell_f.border = THIN_BORDER
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = Font(color="000000")

        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IFERROR(IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row})),\"\")"
        cell_g.number_format = "0.0%"
        cell_g.border = THIN_BORDER
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(color="000000")

    # TOTAL row
    total_row = 6 + len(area_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cell_tot_g = ws.cell(row=total_row, column=7)
    cell_tot_g.value = f"=IFERROR(IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row})),\"\")"
    cell_tot_g.number_format = "0.0%"
    cell_tot_g.font = Font(bold=True, color="000000")
    cell_tot_g.fill = grey_fill
    cell_tot_g.border = THIN_BORDER
    cell_tot_g.alignment = Alignment(horizontal="center")

    # TABLE 2: Key Metrics
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS AND KPIs"
    ws[f"A{metrics_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = navy_fill
    ws[f"A{metrics_start}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[metrics_start].height = 20

    # TABLE 2 header row (GS-SD-016: D9D9D9)
    t2_headers = ["Metric", "Value", "", "", "", "", ""]
    for col, header in enumerate(t2_headers, 1):
        cell = ws.cell(row=metrics_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Ranges match extended sheet data rows:
    # Usage Rules: rows 5-59 | Prohibited Activities: rows 5-59
    # Permitted Activities: rows 5-54 | Handling Requirements: rows 5-54
    metrics = [
        # Usage Rules
        ("Total Usage Rules Documented", "=COUNTA('Usage Rules'!B5:B59)"),
        ("Rules Classified as Permitted", "=COUNTIF('Usage Rules'!D5:D59,\"Permitted\")"),
        ("Rules Permitted with Conditions", "=COUNTIF('Usage Rules'!D5:D59,\"Permitted with Conditions\")"),
        ("Rules Classified as Prohibited", "=COUNTIF('Usage Rules'!D5:D59,\"Prohibited\")"),
        ("Rules with Monitoring Required", "=COUNTIF('Usage Rules'!G5:G59,\"Yes\")"),
        # Permitted Activities
        ("Total Permitted Activities Documented", "=COUNTA('Permitted Activities'!A5:A54)"),
        ("Permitted Activities Requiring Approval", "=COUNTIF('Permitted Activities'!E5:E54,\"Yes\")"),
        # Prohibited Activities
        ("Total Prohibited Activities Documented", "=COUNTA('Prohibited Activities'!A5:A59)"),
        ("Critical-Risk Prohibitions", "=COUNTIF('Prohibited Activities'!E5:E59,\"Critical\")"),
        ("High-Risk Prohibitions", "=COUNTIF('Prohibited Activities'!E5:E59,\"High\")"),
        ("Prohibitions Allowing Exceptions", "=COUNTIF('Prohibited Activities'!H5:H59,\"Yes\")"),
        # Handling Requirements
        ("Total Handling Requirements Defined", "=COUNTA('Handling Requirements'!A5:A54)"),
        ("Requirements for Restricted Data", "=COUNTIF('Handling Requirements'!C5:C54,\"Restricted\")"),
        ("Requirements with Encryption Mandated", "=COUNTIF('Handling Requirements'!I5:I54,\"Yes\")"),
        ("Requirements with Labelling Required", "=COUNTIF('Handling Requirements'!H5:H54,\"Yes\")"),
    ]

    row = metrics_start + 2
    for metric, formula in metrics:
        cell_label = ws.cell(row=row, column=1, value=metric)
        cell_label.border = THIN_BORDER
        cell_label.font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = THIN_BORDER
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TABLE 3: Critical Findings
    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = red_fill
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[crit_start].height = 20

    t3_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, header in enumerate(t3_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # TABLE 3 findings — ranges match extended sheet data rows
    findings = [
        ("Prohibited Activities", "Critical-risk prohibited activities", "=COUNTIF('Prohibited Activities'!E5:E59,\"Critical\")", "Critical", "Verify enforcement"),
        ("Prohibited Activities", "High-risk prohibited activities", "=COUNTIF('Prohibited Activities'!E5:E59,\"High\")", "High", "Add detection controls"),
        ("Prohibited Activities", "Prohibitions allowing exceptions (review needed)", "=COUNTIF('Prohibited Activities'!H5:H59,\"Yes\")", "High", "Review exception process"),
        ("Usage Rules", "Rules without monitoring controls defined", "=COUNTIF('Usage Rules'!G5:G59,\"No\")", "High", "Add monitoring"),
        ("Handling Requirements", "Restricted data assets requiring encryption", "=COUNTIF('Handling Requirements'!I5:I54,\"Yes\")", "Medium", "Verify encryption"),
        ("Handling Requirements", "Assets requiring labelling (verify compliance)", "=COUNTIF('Handling Requirements'!H5:H54,\"Yes\")", "Medium", "Confirm labels applied"),
    ]

    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty FFFFCC buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Column widths and freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register(ws):
    """Create the Evidence Register sheet."""
    hdr_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    col_hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    col_hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    input_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = hdr_font
    ws["A1"].fill = hdr_fill
    ws["A1"].alignment = hdr_align
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files and documents referenced in this assessment (audit traceability)"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Rule", 20),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = col_hdr_font
        cell.fill = col_hdr_fill
        cell.alignment = col_hdr_align
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Policy Document,Configuration,Screenshot,Export,Training Record,Attestation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    # Sample row (row 4) — F2F2F2
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = [
        "EV-001",
        "Policy Document",
        "Acceptable Use Policy v2.3 — usage rules section",
        "UR-001",
        "2026-01-15",
        "/policies/AUP-v2.3.pdf",
        "IT Security Manager",
        "2027-01-15",
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = sample_fill
        cell.border = THIN_BORDER
        cell.alignment = input_align

    # Data entry rows (rows 6-105 = 100 empty rows)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = THIN_BORDER
            cell.alignment = input_align
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws.title = "Approval Sign-Off"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE(\'Summary Dashboard\'!G6:G8),\"\")")  ,
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/8] Creating Usage Rules sheet...")
        create_usage_rules_sheet(wb["Usage Rules"], styles)

        logger.info("[3/8] Creating Permitted Activities sheet...")
        create_permitted_activities_sheet(wb["Permitted Activities"], styles)

        logger.info("[4/8] Creating Prohibited Activities sheet...")
        create_prohibited_activities_sheet(wb["Prohibited Activities"], styles)

        logger.info("[5/8] Creating Handling Requirements sheet...")
        create_handling_requirements_sheet(wb["Handling Requirements"], styles)

        logger.info("[7/8] Creating Evidence Records sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[6/8] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"])

        logger.info("[8/8] Creating Approval Record sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        finalize_validations(wb)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        wb.save(output_path)
        logger.info(f"SUCCESS: {OUTPUT_FILENAME}")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        return 1
    except PermissionError as e:
        logger.error(f"Permission denied: {e}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
