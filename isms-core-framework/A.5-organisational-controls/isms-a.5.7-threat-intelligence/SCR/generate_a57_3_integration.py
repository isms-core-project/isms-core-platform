#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.7.3 - Intelligence Integration & Distribution Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.7: Threat Intelligence
Assessment Domain 3 of 5: Intelligence Integration & Distribution Workflows

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific intelligence distribution requirements, consumer
workflows, and integration architectures.

Key customization areas:
1. Intelligence consumer groups and their requirements (match your org structure)
2. Distribution mechanisms and automation (adapt to your technical stack)
3. MITRE ATT&CK mapping requirements per consumer (align with team needs)
4. Integration platforms and APIs (specific to your security tools)
5. Feedback mechanisms and metrics (based on operational workflows)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.7 Threat Intelligence Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
how analysed threat intelligence (from A.5.7.2) is integrated into security
operations workflows and distributed to intelligence consumers across the
organisation.

**Purpose:**
Enables systematic assessment of intelligence distribution mechanisms, consumer
requirements, feedback loops, and integration effectiveness to ensure threat
intelligence drives actionable security improvements.

**Assessment Scope:**
- Intelligence consumer identification and requirement analysis
- Distribution mechanisms (automated feeds, reports, alerts, portals)
- Integration with security operations workflows (SOC, IR, Vuln Mgmt, IAM)
- MITRE ATT&CK contextualization for different consumer needs
- Indicator distribution to security tools (SIEM, EDR, firewalls, proxies)
- VulnerabilityThreatLink (VTL) workflow integration (A.5.7 ↔ A.8.8)
- Intelligence product types and delivery formats
- Consumer feedback mechanisms and satisfaction metrics
- Distribution timeliness and SLA compliance
- Access control and need-to-know enforcement
- Metrics collection and effectiveness measurement
- Audit evidence collection for compliance validation

**Generated Workbook Structure (13 Sheets):**
1. Instructions - Assessment guidance and integration methodology
2. Consumer_Registry - Intelligence consumers and their requirements
3. Distribution_Mechanisms - Automated and manual distribution methods
4. SOC_Integration - Security Operations Center workflows
5. IR_Integration - Incident Response workflow integration
6. VulnMgmt_Integration - Vulnerability Management (VTL workflow)
7. Tool_Integrations - Security tool indicator deployment
8. Intelligence_Products - Reports, alerts, briefings, dashboards
9. Feedback_Loop - Consumer feedback and satisfaction tracking
10. Performance_Metrics - Distribution SLAs, timeliness, coverage
11. Gap_Analysis - Integration deficiencies and remediation
12. Evidence_Register - Audit evidence tracking
13. Metadata - Assessment metadata and approval workflow

**Key Features:**
- Consumer requirement mapping with MITRE ATT&CK contextualization
- VulnerabilityThreatLink (VTL) workflow integration assessment (CRITICAL)
- Distribution mechanism effectiveness analysis
- Indicator deployment automation tracking
- Consumer satisfaction and feedback metrics
- Data validation with standardized dropdown lists
- Conditional formatting for integration health status
- Protected formulas with unprotected input cells
- Integration with A.5.7.2 (analysis outputs) and consumer workflows
- Evidence linkage for audit traceability

**VulnerabilityThreatLink (VTL) Integration (CRITICAL):**
This assessment specifically tracks the VTL workflow that bridges Control A.5.7
(Threat Intelligence) with Control A.8.8 (Vulnerability Management):

VTL Workflow:
1. Threat intelligence identifies active exploitation (A.5.7.2 analysis)
2. CVSS-scored vulnerability data correlated with exploitation intelligence
3. Automated emergency patching triggered for exploited vulnerabilities
4. Vulnerability management team receives prioritized remediation queue
5. Feedback on remediation success returns to threat intelligence

This bi-directional integration is assessed in Sheet 6 (VulnMgmt_Integration)
and is AUDIT CRITICAL for demonstrating operational security value.

**Integration:**
This assessment feeds into:
- A.5.7.4 Threat Intelligence Effectiveness Dashboard (distribution metrics)
- A.5.7.2 Collection & Analysis (consumer feedback for requirements)
- A.8.8 Vulnerability Management (VTL workflow effectiveness)
- Consumer workflows: SOC, IR, Vuln Mgmt, IAM, Network Security

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a57_3_integration.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a57_3_integration.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a57_3_integration.py --date 20250115

Output:
    File: ISMS_A_5_7_3_Integration_Distribution_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)
    Sheets: 13 worksheets (see structure above)

Post-Generation Steps:
    1. Identify all threat intelligence consumers (teams, tools, processes)
    2. Document consumer requirements (format, frequency, MITRE mapping needs)
    3. Map distribution mechanisms (automated feeds, reports, portals, APIs)
    4. Assess SOC integration (indicator deployment, alert enrichment, hunting)
    5. Evaluate IR integration (threat context for incidents, playbook triggers)
    6. Validate VulnMgmt integration (VTL workflow - CRITICAL)
    7. Review tool integrations (SIEM, EDR, firewall, proxy IOC deployment)
    8. Catalog intelligence products (types, formats, audiences)
    9. Implement feedback mechanisms (consumer surveys, effectiveness metrics)
    10. Measure distribution performance (SLAs, timeliness, accuracy)
    11. Identify integration gaps and create remediation plans
    12. Collect audit evidence (distribution logs, feedback data, metrics)
    13. Obtain stakeholder approvals
    14. Feed results into A.5.7.4 Effectiveness Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.7
Assessment Domain:    3 of 5 (Intelligence Integration & Distribution Workflows)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.7: Threat Intelligence Policy (Governance)
    - ISMS-IMP-A.5.7.3: Intelligence Integration & Distribution Implementation Guide
    - ISMS-IMP-A.5.7.1: Threat Intelligence Sources Assessment (Domain 1)
    - ISMS-IMP-A.5.7.2: Intelligence Collection & Analysis Assessment (Domain 2)
    - ISMS-POL-A.8.8: Vulnerability Management Policy (VTL integration)
    - ISMS-IMP-A.8.8: Vulnerability Management Implementation (VTL workflow)
    - ISO 27002:2022 Implementation Guidance for Control A.5.7
    - MITRE ATT&CK Framework (Enterprise v15.1)
    - STIX/TAXII Standards for Threat Intelligence Exchange

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
- Added dedicated VulnMgmt_Integration sheet (Sheet 6) for VTL workflow
    - Expanded consumer requirement analysis with MITRE ATT&CK mapping
    - Enhanced feedback loop mechanisms and satisfaction metrics
    - Added intelligence product catalog and format standardization
    - Improved tool integration tracking (SIEM, EDR, firewall, proxy)
    - Added distribution timeliness and SLA compliance metrics
    - Enhanced audit evidence collection mechanisms
    - Updated conditional formatting for integration health visualization

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Distribution Effectiveness Standards:**
Threat intelligence value is realised only when consumed operationally.
Establish clear effectiveness metrics:
- Consumer reach: Percentage of eligible consumers receiving intelligence
- Timeliness: Time from analysis completion to consumer delivery
- Actionability: Percentage of distributed intelligence leading to security actions
- Satisfaction: Consumer feedback scores and improvement requests
- Coverage: Percentage of consumer requirements met by intelligence products

Poor distribution effectiveness wastes threat intelligence investments.

**VulnerabilityThreatLink (VTL) Workflow (AUDIT CRITICAL):**
The VTL workflow is the most critical integration assessed in this workbook.
Auditors will specifically verify:
- Automated correlation between threat intelligence and vulnerability data
- Emergency patching trigger mechanisms for actively exploited CVEs
- Bi-directional data flow (A.5.7 → A.8.8 → A.5.7 feedback)
- CVSS accuracy validation (from A.5.7.1 and A.5.7.2)
- Remediation effectiveness tracking and feedback loops

Complete Sheet 6 (VulnMgmt_Integration) with comprehensive detail. This
integration demonstrates operational security value beyond compliance theater.

**Consumer Requirement Diversity:**
Different consumers need intelligence in different formats:
- SOC analysts: Real-time IOC feeds, SIEM correlation rules, hunting queries
- Incident responders: Threat actor TTPs, campaign context, attribution
- Vulnerability managers: Exploited CVE lists, CVSS scores, patch priorities
- Network security: IP/domain blocklists, malicious infrastructure data
- Executives: Strategic intelligence briefings, risk trends, threat landscape

Tailor distribution mechanisms to consumer needs, not one-size-fits-all.

**MITRE ATT&CK Contextualization:**
Intelligence consumers benefit from ATT&CK contextualization:
- SOC: Techniques for detection engineering and threat hunting
- IR: Techniques for incident scoping and lateral movement analysis
- Red Team: Techniques for adversary emulation exercises
- Security Architecture: Techniques for control gap analysis

Map intelligence products to relevant ATT&CK techniques per consumer needs.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will verify:
- Documented distribution workflows (automated and manual)
- Consumer requirement documentation and satisfaction tracking
- Integration with operational security workflows (SOC, IR, Vuln Mgmt)
- VTL workflow implementation and effectiveness (CRITICAL)
- Feedback mechanisms and continuous improvement

Maintain evidence in Sheet 12 (Evidence_Register) with proper linkage.

**Data Protection:**
Assessment workbooks contain sensitive operational details:
- Intelligence consumer identities and organisational structure
- Integration technical details (APIs, credentials, endpoints)
- Security tool configurations and IOC deployment mechanisms
- Performance metrics revealing operational gaps

Classification: CONFIDENTIAL - Internal Use Only. Restrict to security team.

**Maintenance:**
Review and update assessment:
- Monthly: Distribution performance metrics and SLA compliance
- Quarterly: Consumer satisfaction surveys and requirement updates
- Semi-annually: Tool integration review and optimization
- Annually: Complete reassessment of distribution effectiveness
- Ad-hoc: When new consumers identified or workflows change

**Quality Assurance:**
Have threat intelligence team leads, SOC management, and vulnerability
management leads validate assessments before using results for workflow
changes or resource allocation. VTL workflow effectiveness must be
cross-validated with Control A.8.8 assessment data.

**Regulatory Alignment:**
Some regulations require specific intelligence distribution controls:
- Financial sector: Fraud intelligence sharing with industry groups
- Healthcare: Protected health information considerations in intelligence sharing
- Critical infrastructure: Information sharing with government agencies (e.g., CISA)
- Data protection: GDPR/privacy considerations for threat actor attribution

Customize distribution assessments to include regulatory-specific requirements.

**Indicator Deployment Automation:**
Manual indicator deployment to security tools doesn't scale. Prioritize:
- STIX/TAXII automated feeds to TIP/SIEM
- API-based IOC deployment to EDR, firewalls, proxies
- Automated blocklist updates based on confidence thresholds
- Orchestrated indicator expiration and removal

Track automation coverage in Sheet 7 (Tool_Integrations).

**Feedback Loop Importance:**
Consumer feedback drives continuous improvement:
- Intelligence gaps: Threats not covered by current sources/analysis
- Format issues: Intelligence products not matching workflow needs
- Timeliness problems: Intelligence arriving too late for action
- False positives: Indicators causing operational friction
- Success stories: Intelligence enabling successful threat detection/response

Establish formal feedback mechanisms (Sheet 9) and act on feedback quarterly.

**Business Impact:**
Poor intelligence integration and distribution results in:
- Threat intelligence remaining unused (wasted investment)
- Delayed threat detection and incident response
- Missed opportunities for proactive vulnerability remediation
- Analyst frustration with irrelevant or poorly formatted intelligence
- Compliance gaps (intelligence available but not operationally integrated)

This assessment helps identify and remediate distribution/integration gaps.

**VTL Workflow Example:**
Example VTL workflow that should be documented in Sheet 6:

1. Threat intel identifies active exploitation of CVE-2024-XXXXX (A.5.7.2)
2. CVSS 4.0 score: 9.3 Critical with exploit maturity: High
3. Automated correlation with vulnerability scan data (A.8.8)
4. Emergency patch workflow triggered if vulnerable systems identified
5. Vulnerability team receives prioritized remediation ticket
6. Remediation tracked with 24-hour SLA for critical exploited CVEs
7. Feedback to threat intel on remediation success and any indicators observed

This end-to-end workflow must be documented, tested, and measured.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.7.3"
WORKBOOK_NAME = "Intelligence Integration & Distribution Assessment"
CONTROL_ID = "A.5.7"
CONTROL_NAME = "Threat Intelligence"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all 17 required sheets matching specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure for ISMS-IMP-A.5.7.3 (12 sheets — Gold Standard tail order: data → ER → SD → AS)
    sheets = [
        "Instructions & Legend",
        "Tool Integration Matrix",
        "IOC Deployment",
        "Dissemination Channels",
        "Stakeholder Registry",
        "Distribution Tracking",
        "Feedback Collection",
        "Integration Metrics",
        "Action Items",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles



_STYLES = setup_styles()
def get_border():
    """Create new border object (avoid shared object warnings)."""
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATION FUNCTIONS
# ============================================================================

def setup_validations():
    """Create all data validation dropdowns."""
    validations = {}

    # Tool Category
    validations['tool_category'] = DataValidation(
        type="list",
        formula1='"SIEM,EDR,Firewall,Proxy,Email_Gateway,IDS_IPS,TIP,SOAR,Vuln_Scanner,Other"',
        allow_blank=False
    )
    
    # Integration Status
    validations['integration_status'] = DataValidation(
        type="list",
        formula1='"Fully_Integrated,Partially_Integrated,Planned,Not_Integrated"',
        allow_blank=False
    )
    
    # Integration Method
    validations['integration_method'] = DataValidation(
        type="list",
        formula1='"API,Feed,Manual_Import,STIX_TAXII,Syslog,Custom_Script"',
        allow_blank=False
    )
    
    # Integration Direction
    validations['integration_direction'] = DataValidation(
        type="list",
        formula1='"Inbound,Outbound,Bidirectional"',
        allow_blank=False
    )
    
    # Automation Level
    validations['automation_level'] = DataValidation(
        type="list",
        formula1='"Fully_Automated,Semi_Automated,Manual"',
        allow_blank=False
    )
    
    # Update Frequency
    validations['update_frequency'] = DataValidation(
        type="list",
        formula1='"Real-time,Hourly,Daily,Weekly,On_Demand"',
        allow_blank=False
    )
    
    # IOC Type
    validations['ioc_type'] = DataValidation(
        type="list",
        formula1='"IP_Address,Domain,URL,File_Hash,Email_Address,Registry_Key,Mutex,Certificate"',
        allow_blank=False
    )
    
    # Confidence Level
    validations['confidence'] = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=False
    )
    
    # Severity
    validations['severity'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,Info"',
        allow_blank=False
    )
    
    # TLP Classification
    validations['tlp'] = DataValidation(
        type="list",
        formula1='"TLP:CLEAR,TLP:GREEN,TLP:AMBER,TLP:AMBER+STRICT,TLP:RED"',
        allow_blank=False
    )
    
    # IOC Status
    validations['ioc_status'] = DataValidation(
        type="list",
        formula1='"Active,Expired,Withdrawn,Under_Review"',
        allow_blank=False
    )
    
    # Yes/No/Unknown
    validations['yes_no_unk'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    
    # False Positive Status
    validations['false_positive'] = DataValidation(
        type="list",
        formula1='"Yes,No,Under_Investigation"',
        allow_blank=True
    )
    
    # Rating (1-5)
    validations['rating'] = DataValidation(
        type="list",
        formula1='"Excellent,Good,Fair,Poor"',
        allow_blank=True
    )
    
    # Priority
    validations['priority'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    
    # Action Status
    validations['action_status'] = DataValidation(
        type="list",
        formula1='"Open,In_Progress,Blocked,Resolved,Closed"',
        allow_blank=False
    )
    
    return validations


# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS
# ============================================================================

def create_instructions(ws, styles):
    """Create GS-IL-compliant Instructions & Legend sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 2 — empty

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Workbook Title", WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Instructions section
    row = 12
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Document all security tools in scope for threat intelligence integration in the Tool Integration Matrix sheet (SIEM, SOAR, IDS/IPS, firewall, EDR, vulnerability scanners).",
        "2. Assess each tool's integration status, feed ingestion capability, and supported data formats (STIX/TAXII, CSV, JSON, XML, MISP).",
        "3. Record integration performance metrics in the Integration Metrics sheet — track timeliness, volume ingested, false positive rate, and analyst time saved.",
        "4. Identify tools lacking active threat intelligence integration and document gap remediation plans with target completion dates and responsible owners.",
        "5. Use the Performance Trend column to track whether each integration metric is Improving, Stable, or Declining quarter-on-quarter.",
        "6. Prioritise automation for high-volume manual processes — document automation opportunities and measure analyst workload reduction.",
        "7. Maintain the Evidence Register with integration configuration exports, API documentation, and performance dashboards.",
        "8. Obtain sign-off from the SOC Manager, Security Architecture lead, and CISO.",
    ]
    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # Status Legend section
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("✓", "Compliant / Complete", "Requirement fully met", _green),
        ("⚠", "Partial / In Progress", "Partially met or in progress", _amber),
        ("✗", "Non-Compliant / Not Started", "Requirement not met", _red),
        ("—", "Not Applicable", "Not applicable to this assessment", None),
    ]
    row += 1
    for sym, status, desc, fill in legend_rows:
        ws.cell(row=row, column=1, value=sym).border = _border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    # Acceptable Evidence section
    row += 1
    ws[f"A{row}"] = "Acceptable Evidence (examples)"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for evidence in [
        "\u2713 SIEM / SOAR threat intelligence integration configuration exports (sanitised)",
        "\u2713 STIX / TAXII and API integration documentation",
        "\u2713 Integration performance reports and dashboards",
        "\u2713 Automation workflow documentation and efficiency metrics",
        "\u2713 Vendor integration capability documentation and data sheets",
    ]:
        ws[f"A{row}"] = evidence
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"
def create_tool_integration(ws, styles, validations):
    """Create Tool_Integration_Matrix sheet - security tool integration status."""
    
    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "SECURITY TOOL INTEGRATION MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Document threat intelligence integration status across all security tools - SIEM, EDR, firewalls, proxies, email gateways, etc."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Tool ID",
        "Tool Category",
        "Tool Name",
        "Vendor",
        "Version",
        "Primary Owner",
        "Integration Status",
        "Integration Method",
        "Integration Direction",
        "Data Types Integrated",
        "Automation Level",
        "Update Frequency",
        "IOC Types Supported",
        "Last Successful Sync",
        "Sync Errors Last 30 Days",
        "Effectiveness Rating",
    ]
    
    col_widths = [12, 15, 30, 20, 12, 20, 18, 18, 18, 25, 18, 18, 30, 20, 22, 18]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Data rows (30 tools capacity)
    for row in range(5, 35):
        # Tool_ID (auto-generated format)
        if row == 5:
            ws[f"A{row}"] = "TOOL-001"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # Tool_Category dropdown (B)
        validations['tool_category'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Tool_Name, Vendor, Version, Primary_Owner (C-F)
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
        
        # Integration_Status dropdown (G)
        validations['integration_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        
        # Integration_Method dropdown (H)
        validations['integration_method'].add(ws[f"H{row}"])
        ws[f"H{row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        
        # Integration_Direction dropdown (I)
        validations['integration_direction'].add(ws[f"I{row}"])
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        
        # Data_Types_Integrated (J)
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        
        # Automation_Level dropdown (K)
        validations['automation_level'].add(ws[f"K{row}"])
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()
        
        # Update_Frequency dropdown (L)
        validations['update_frequency'].add(ws[f"L{row}"])
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()
        
        # IOC_Types_Supported (M)
        ws[f"M{row}"].fill = styles["input_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        
        # Last_Successful_Sync (N) - datetime
        ws[f"N{row}"].fill = styles["input_cell"]["fill"]
        ws[f"N{row}"].border = get_border()
        ws[f"N{row}"].number_format = 'DD.MM.YYYY hh:mm'
        
        # Sync_Errors_Last_30_Days (O) - number
        ws[f"O{row}"].fill = styles["input_cell"]["fill"]
        ws[f"O{row}"].border = get_border()
        ws[f"O{row}"].number_format = '0'
        
        # Effectiveness_Rating dropdown (P)
        validations['rating'].add(ws[f"P{row}"])
        ws[f"P{row}"].fill = styles["input_cell"]["fill"]
        ws[f"P{row}"].border = get_border()
    
    # Apply data validations
    ws.add_data_validation(validations['tool_category'])
    ws.add_data_validation(validations['integration_status'])
    ws.add_data_validation(validations['integration_method'])
    ws.add_data_validation(validations['integration_direction'])
    ws.add_data_validation(validations['automation_level'])
    ws.add_data_validation(validations['update_frequency'])
    ws.add_data_validation(validations['rating'])
    
    # Conditional formatting
    # Integration_Status "Fully_Integrated" → Green
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    for row in range(5, 35):
        rule1 = CellIsRule(operator='equal', formula=['"Fully_Integrated"'], fill=green_fill)
        ws.conditional_formatting.add(f"G{row}", rule1)
        
        rule2 = CellIsRule(operator='equal', formula=['"Not_Integrated"'], fill=red_fill)
        ws.conditional_formatting.add(f"G{row}", rule2)
    
    # Sync_Errors > 5 → Yellow
    for row in range(5, 35):
        rule = CellIsRule(operator='greaterThan', formula=['5'], fill=yellow_fill)
        ws.conditional_formatting.add(f"O{row}", rule)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    row = 36
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "INTEGRATION SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Total Tools:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(C5:C34)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Fully Integrated:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(G5:G34,"Fully_Integrated")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "Partially Integrated:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(G5:G34,"Partially_Integrated")'
    ws[f"B{row}"].font = Font(bold=True, color="C65911")
    
    row += 1
    ws[f"A{row}"] = "Not Integrated:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(G5:G34,"Not_Integrated")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")


# ============================================================================
# SECTION 5: SHEET 3 - IOC_DEPLOYMENT
# ============================================================================

def create_ioc_deployment(ws, styles, validations):
    """Create IOC_Deployment sheet - IOC tracking and effectiveness."""
    
    # Title
    ws.merge_cells("A1:T1")
    ws["A1"] = "INDICATOR OF COMPROMISE (IOC) DEPLOYMENT TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:T2")
    ws["A2"] = "Track IOC deployment lifecycle, detection effectiveness, and false positive management. Correlate IOCs to incidents and vulnerabilities."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "IOC ID",
        "IOC Type",
        "IOC Value",
        "Intelligence Source",
        "Threat Actor",
        "Associated Malware",
        "Confidence Level",
        "Severity",
        "TLP Classification",
        "Deployment Date",
        "Deployed To Tools",
        "Deployment Method",
        "Expiration Date",
        "Status",
        "Hits Last 7 Days",
        "Hits Last 30 Days",
        "Hits Total",
        "Last Hit Date",
        "False Positive",
        "Effectiveness Rating",
    ]
    
    col_widths = [15, 15, 30, 18, 18, 20, 18, 12, 20, 15, 25, 18, 15, 15, 18, 18, 12, 15, 15, 18]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Data rows (100 IOCs capacity)
    for row in range(5, 105):
        # IOC_ID (auto-generated format)
        if row == 5:
            ws[f"A{row}"] = "IOC-2025-001"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # IOC_Type dropdown (B)
        validations['ioc_type'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # IOC_Value (C)
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        # Intelligence_Source, Threat_Actor, Associated_Malware (D-F)
        for col in ["D", "E", "F"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
        
        # Confidence_Level dropdown (G)
        validations['confidence'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        
        # Severity dropdown (H)
        validations['severity'].add(ws[f"H{row}"])
        ws[f"H{row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        
        # TLP_Classification dropdown (I)
        validations['tlp'].add(ws[f"I{row}"])
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        
        # Deployment_Date (J)
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        ws[f"J{row}"].number_format = 'DD.MM.YYYY'
        
        # Deployed_To_Tools (K)
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()
        
        # Deployment_Method dropdown (L)
        deploy_method_dv = DataValidation(
            type="list",
            formula1='"Automated,Manual,Hybrid"',
            allow_blank=False
        )
        deploy_method_dv.add(ws[f"L{row}"])
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()
        
        # Expiration_Date (M)
        ws[f"M{row}"].fill = styles["input_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        ws[f"M{row}"].number_format = 'DD.MM.YYYY'
        
        # Status dropdown (N)
        validations['ioc_status'].add(ws[f"N{row}"])
        ws[f"N{row}"].fill = styles["input_cell"]["fill"]
        ws[f"N{row}"].border = get_border()
        
        # Hits counts (O-Q) - numbers
        for col in ["O", "P", "Q"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
            ws[f"{col}{row}"].number_format = '0'
        
        # Last_Hit_Date (R)
        ws[f"R{row}"].fill = styles["input_cell"]["fill"]
        ws[f"R{row}"].border = get_border()
        ws[f"R{row}"].number_format = 'DD.MM.YYYY'
        
        # False_Positive dropdown (S)
        validations['false_positive'].add(ws[f"S{row}"])
        ws[f"S{row}"].fill = styles["input_cell"]["fill"]
        ws[f"S{row}"].border = get_border()
        
        # Effectiveness_Rating formula (T)
        ws[f"T{row}"] = f'=IF(AND(Q{row}>0,S{row}="No"),"Effective",IF(AND(Q{row}>0,S{row}="Yes"),"Ineffective",IF(AND(Q{row}=0,J{row}<TODAY()-30),"Ineffective","Monitoring")))'
        ws[f"T{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"T{row}"].border = get_border()
    
    # Apply custom data validation for deployment method
    for row in range(5, 105):
        deploy_method_dv = DataValidation(
            type="list",
            formula1='"Automated,Manual,Hybrid"',
            allow_blank=False
        )
        ws.add_data_validation(deploy_method_dv)
        deploy_method_dv.add(ws[f"L{row}"])
    
    # Apply other validations
    ws.add_data_validation(validations['ioc_type'])
    ws.add_data_validation(validations['confidence'])
    ws.add_data_validation(validations['severity'])
    ws.add_data_validation(validations['tlp'])
    ws.add_data_validation(validations['ioc_status'])
    ws.add_data_validation(validations['false_positive'])
    
    # Conditional formatting
    # Status "Expired" → Gray
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    for row in range(5, 105):
        rule1 = CellIsRule(operator='equal', formula=['"Expired"'], fill=gray_fill)
        ws.conditional_formatting.add(f"A{row}:T{row}", rule1)
        
        # Hits_Last_7_Days > 0 → Green
        rule2 = CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
        ws.conditional_formatting.add(f"O{row}", rule2)
        
        # False_Positive "Yes" → Red
        rule3 = CellIsRule(operator='equal', formula=['"Yes"'], fill=red_fill)
        ws.conditional_formatting.add(f"S{row}", rule3)
        
        # Effectiveness "Ineffective" → Red text
        rule4 = CellIsRule(operator='equal', formula=['"Ineffective"'], fill=red_fill)
        ws.conditional_formatting.add(f"T{row}", rule4)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    row = 106
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "IOC SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Total IOCs:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(C5:C104)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Active IOCs:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(N5:N104,"Active")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "IOCs with Hits (7 days):"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(O5:O104,">0")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "False Positives:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(S5:S104,"Yes")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")
    
    row += 1
    ws[f"A{row}"] = "Effective IOCs:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(T5:T104,"Effective")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")

# ============================================================================
# SECTION 6: SHEET 4 - DISSEMINATION_CHANNELS
# ============================================================================

def create_dissemination_channels(ws, styles, validations):
    """Create Dissemination_Channels sheet - intelligence distribution mechanisms."""
    
    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "INTELLIGENCE DISSEMINATION CHANNELS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:N2")
    ws["A2"] = "Document all channels used to distribute threat intelligence - email lists, portals, dashboards, APIs, briefings, chat platforms."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Channel ID",
        "Channel Name",
        "Channel Type",
        "Target Audience",
        "Intelligence Types",
        "Frequency",
        "Delivery Method",
        "TLP Max Classification",
        "Active Subscribers",
        "Avg Engagement Rate",
        "Status",
        "Owner",
        "Last Distribution",
        "Notes",
    ]
    
    col_widths = [12, 25, 20, 20, 25, 15, 20, 22, 18, 20, 12, 20, 18, 35]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Data rows (25 channels capacity)
    for row in range(5, 30):
        # Channel_ID (auto-generated format)
        if row == 5:
            ws[f"A{row}"] = "CHAN-001"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # Channel_Name (B)
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Channel_Type dropdown (C)
        channel_type_dv = DataValidation(
            type="list",
            formula1='"Email_List,Portal,Dashboard,API,Briefing,Chat,Ticketing_System"',
            allow_blank=False
        )
        channel_type_dv.add(ws[f"C{row}"])
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        # Target_Audience dropdown (D)
        audience_dv = DataValidation(
            type="list",
            formula1='"Executive,CISO,SOC,IR_Team,IT_Ops,Developers,All_Staff"',
            allow_blank=False
        )
        audience_dv.add(ws[f"D{row}"])
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        # Intelligence_Types (E)
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = get_border()
        
        # Frequency dropdown (F)
        frequency_dv = DataValidation(
            type="list",
            formula1='"Real-time,Daily,Weekly,Monthly,On_Demand,Quarterly"',
            allow_blank=False
        )
        frequency_dv.add(ws[f"F{row}"])
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        
        # Delivery_Method dropdown (G)
        delivery_dv = DataValidation(
            type="list",
            formula1='"Email,Web_Portal,Mobile_App,API,PDF_Report,Presentation,Slack,Teams"',
            allow_blank=False
        )
        delivery_dv.add(ws[f"G{row}"])
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        
        # TLP_Max_Classification dropdown (H)
        validations['tlp'].add(ws[f"H{row}"])
        ws[f"H{row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        
        # Active_Subscribers (I) - number
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        ws[f"I{row}"].number_format = '0'
        
        # Avg_Engagement_Rate (J) - percentage
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        ws[f"J{row}"].number_format = '0"%"'
        
        # Status dropdown (K)
        status_dv = DataValidation(
            type="list",
            formula1='"Active,Inactive,Pilot,Deprecated"',
            allow_blank=False
        )
        status_dv.add(ws[f"K{row}"])
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()
        
        # Owner (L)
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()
        
        # Last_Distribution (M)
        ws[f"M{row}"].fill = styles["input_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        ws[f"M{row}"].number_format = 'DD.MM.YYYY'
        
        # Notes (N)
        ws[f"N{row}"].fill = styles["input_cell"]["fill"]
        ws[f"N{row}"].border = get_border()
    
    # Apply custom data validations
    for row in range(5, 30):
        channel_type_dv = DataValidation(
            type="list",
            formula1='"Email_List,Portal,Dashboard,API,Briefing,Chat,Ticketing_System"',
            allow_blank=False
        )
        ws.add_data_validation(channel_type_dv)
        channel_type_dv.add(ws[f"C{row}"])
        
        audience_dv = DataValidation(
            type="list",
            formula1='"Executive,CISO,SOC,IR_Team,IT_Ops,Developers,All_Staff"',
            allow_blank=False
        )
        ws.add_data_validation(audience_dv)
        audience_dv.add(ws[f"D{row}"])
        
        frequency_dv = DataValidation(
            type="list",
            formula1='"Real-time,Daily,Weekly,Monthly,On_Demand,Quarterly"',
            allow_blank=False
        )
        ws.add_data_validation(frequency_dv)
        frequency_dv.add(ws[f"F{row}"])
        
        delivery_dv = DataValidation(
            type="list",
            formula1='"Email,Web_Portal,Mobile_App,API,PDF_Report,Presentation,Slack,Teams"',
            allow_blank=False
        )
        ws.add_data_validation(delivery_dv)
        delivery_dv.add(ws[f"G{row}"])
        
        status_dv = DataValidation(
            type="list",
            formula1='"Active,Inactive,Pilot,Deprecated"',
            allow_blank=False
        )
        ws.add_data_validation(status_dv)
        status_dv.add(ws[f"K{row}"])
    
    # Apply TLP validation
    ws.add_data_validation(validations['tlp'])
    
    # Conditional formatting
    # Status "Active" → Green
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    for row in range(5, 30):
        rule1 = CellIsRule(operator='equal', formula=['"Active"'], fill=green_fill)
        ws.conditional_formatting.add(f"K{row}", rule1)
        
        rule2 = CellIsRule(operator='equal', formula=['"Inactive"'], fill=gray_fill)
        ws.conditional_formatting.add(f"K{row}", rule2)
        
        rule3 = CellIsRule(operator='equal', formula=['"Deprecated"'], fill=gray_fill)
        ws.conditional_formatting.add(f"K{row}", rule3)
        
        rule4 = CellIsRule(operator='equal', formula=['"Pilot"'], fill=yellow_fill)
        ws.conditional_formatting.add(f"K{row}", rule4)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    row = 31
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "CHANNEL SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Total Channels:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(B5:B29)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Active Channels:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(K5:K29,"Active")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "Total Subscribers:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=SUM(I5:I29)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Average Engagement:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=AVERAGE(J5:J29)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '0"%"'


# ============================================================================
# SECTION 7: SHEET 5 - STAKEHOLDER_REGISTRY
# ============================================================================

def create_stakeholder_registry(ws, styles, validations):
    """Create Stakeholder_Registry sheet - audience management."""
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "THREAT INTELLIGENCE STAKEHOLDER REGISTRY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Maintain registry of all threat intelligence stakeholders - track roles, clearances, engagement levels, and communication preferences."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Stakeholder ID",
        "Full Name",
        "Email",
        "Role Title",
        "Department",
        "Stakeholder Type",
        "TLP Clearance",
        "Subscribed Channels",
        "Engagement Level",
        "Last Feedback Date",
        "Status",
        "Notes",
    ]
    
    col_widths = [15, 25, 30, 25, 20, 20, 20, 30, 18, 20, 12, 35]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Data rows (50 stakeholders capacity)
    for row in range(5, 55):
        # Stakeholder_ID (auto-generated format)
        if row == 5:
            ws[f"A{row}"] = "STK-001"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # Full_Name, Email, Role_Title, Department (B-E)
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
        
        # Stakeholder_Type dropdown (F)
        stakeholder_type_dv = DataValidation(
            type="list",
            formula1='"Executive,Manager,Analyst,Operator,Developer,External_Partner,Vendor"',
            allow_blank=False
        )
        stakeholder_type_dv.add(ws[f"F{row}"])
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        
        # TLP_Clearance dropdown (G)
        validations['tlp'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        
        # Subscribed_Channels (H)
        ws[f"H{row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        
        # Engagement_Level dropdown (I)
        engagement_dv = DataValidation(
            type="list",
            formula1='"High,Medium,Low,None"',
            allow_blank=False
        )
        engagement_dv.add(ws[f"I{row}"])
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        
        # Last_Feedback_Date (J)
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        ws[f"J{row}"].number_format = 'DD.MM.YYYY'
        
        # Status dropdown (K)
        status_dv = DataValidation(
            type="list",
            formula1='"Active,Inactive,On_Leave,Departed"',
            allow_blank=False
        )
        status_dv.add(ws[f"K{row}"])
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()
        
        # Notes (L)
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()
    
    # Apply custom data validations
    for row in range(5, 55):
        stakeholder_type_dv = DataValidation(
            type="list",
            formula1='"Executive,Manager,Analyst,Operator,Developer,External_Partner,Vendor"',
            allow_blank=False
        )
        ws.add_data_validation(stakeholder_type_dv)
        stakeholder_type_dv.add(ws[f"F{row}"])
        
        engagement_dv = DataValidation(
            type="list",
            formula1='"High,Medium,Low,None"',
            allow_blank=False
        )
        ws.add_data_validation(engagement_dv)
        engagement_dv.add(ws[f"I{row}"])
        
        status_dv = DataValidation(
            type="list",
            formula1='"Active,Inactive,On_Leave,Departed"',
            allow_blank=False
        )
        ws.add_data_validation(status_dv)
        status_dv.add(ws[f"K{row}"])
    
    # Apply TLP validation
    ws.add_data_validation(validations['tlp'])
    
    # Conditional formatting
    # Engagement_Level colors
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    none_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    for row in range(5, 55):
        rule1 = CellIsRule(operator='equal', formula=['"High"'], fill=high_fill)
        ws.conditional_formatting.add(f"I{row}", rule1)
        
        rule2 = CellIsRule(operator='equal', formula=['"Medium"'], fill=medium_fill)
        ws.conditional_formatting.add(f"I{row}", rule2)
        
        rule3 = CellIsRule(operator='equal', formula=['"Low"'], fill=low_fill)
        ws.conditional_formatting.add(f"I{row}", rule3)
        
        rule4 = CellIsRule(operator='equal', formula=['"None"'], fill=none_fill)
        ws.conditional_formatting.add(f"I{row}", rule4)
    
    # Status "Active" → Green
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    for row in range(5, 55):
        rule1 = CellIsRule(operator='equal', formula=['"Active"'], fill=green_fill)
        ws.conditional_formatting.add(f"K{row}", rule1)
        
        rule2 = CellIsRule(operator='equal', formula=['"Inactive"'], fill=gray_fill)
        ws.conditional_formatting.add(f"K{row}", rule2)
        
        rule3 = CellIsRule(operator='equal', formula=['"Departed"'], fill=gray_fill)
        ws.conditional_formatting.add(f"K{row}", rule3)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    row = 56
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "STAKEHOLDER SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Total Stakeholders:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(B5:B54)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Active Stakeholders:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(K5:K54,"Active")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "High Engagement:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(I5:I54,"High")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "Low/No Engagement:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(I5:I54,"Low")+COUNTIF(I5:I54,"None")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")
    
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "CLEARANCE BREAKDOWN"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "TLP:RED:"
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"TLP:RED")'
    row += 1
    ws[f"A{row}"] = "TLP:AMBER+STRICT:"
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"TLP:AMBER+STRICT")'
    row += 1
    ws[f"A{row}"] = "TLP:AMBER:"
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"TLP:AMBER")'
    row += 1
    ws[f"A{row}"] = "TLP:GREEN:"
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"TLP:GREEN")'
    row += 1
    ws[f"A{row}"] = "TLP:CLEAR:"
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"TLP:CLEAR")'

# ============================================================================
# SECTION 8: SHEET 6 - DISTRIBUTION_TRACKING
# ============================================================================

def create_distribution_tracking(ws, styles, validations):
    """Create Distribution_Tracking sheet - intelligence distribution records."""
    
    # Title
    ws.merge_cells("A1:O1")
    ws["A1"] = "INTELLIGENCE DISTRIBUTION TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:O2")
    ws["A2"] = "Track who received what intelligence, when, and through which channel. Monitor engagement and action rates."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Distribution ID",
        "Product ID",
        "Product Title",
        "Channel ID",
        "Channel Name",
        "Distribution Date",
        "Recipients Count",
        "TLP Classification",
        "Views Count",
        "Downloads Count",
        "Feedback Received",
        "Action Taken Count",
        "Engagement Rate",
        "Effectiveness",
        "Notes",
    ]
    
    col_widths = [15, 15, 30, 12, 25, 18, 18, 20, 15, 18, 18, 18, 18, 15, 30]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Data rows (100 distributions capacity)
    for row in range(5, 105):
        # Distribution_ID (auto-generated format)
        if row == 5:
            ws[f"A{row}"] = "DIST-2025-001"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # Product_ID (B)
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Product_Title (C) - can be manual or VLOOKUP from 5.7.2
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        # Channel_ID (D)
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        # Channel_Name (E) - formula from Dissemination_Channels
        ws[f"E{row}"] = "=IFERROR(VLOOKUP(D" + str(row) + ",\'Dissemination Channels\'!A:B,2,FALSE),\"\")"
        ws[f"E{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"E{row}"].border = get_border()
        
        # Distribution_Date (F)
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        ws[f"F{row}"].number_format = 'DD.MM.YYYY'
        
        # Recipients_Count (G)
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        ws[f"G{row}"].number_format = '0'
        
        # TLP_Classification dropdown (H)
        validations['tlp'].add(ws[f"H{row}"])
        ws[f"H{row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        
        # Views_Count (I)
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        ws[f"I{row}"].number_format = '0'
        
        # Downloads_Count (J)
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        ws[f"J{row}"].number_format = '0'
        
        # Feedback_Received (K)
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()
        ws[f"K{row}"].number_format = '0'
        
        # Action_Taken_Count (L)
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()
        ws[f"L{row}"].number_format = '0'
        
        # Engagement_Rate formula (M) = (Views + Downloads + Feedback) / Recipients
        ws[f"M{row}"] = f'=IF(G{row}=0,"",((I{row}+J{row}+K{row})/G{row})*100)'
        ws[f"M{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        ws[f"M{row}"].number_format = '0.0"%"'
        
        # Effectiveness formula (N)
        ws[f"N{row}"] = f'=IF(AND(L{row}/G{row}>0.25,M{row}>40),"High",IF(OR(L{row}/G{row}>0.15,M{row}>25),"Medium","Low"))'
        ws[f"N{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"N{row}"].border = get_border()
        
        # Notes (O)
        ws[f"O{row}"].fill = styles["input_cell"]["fill"]
        ws[f"O{row}"].border = get_border()
    
    # Apply TLP validation
    ws.add_data_validation(validations['tlp'])
    
    # Conditional formatting
    # Effectiveness colors
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in range(5, 105):
        rule1 = CellIsRule(operator='equal', formula=['"High"'], fill=high_fill)
        ws.conditional_formatting.add(f"N{row}", rule1)
        
        rule2 = CellIsRule(operator='equal', formula=['"Medium"'], fill=medium_fill)
        ws.conditional_formatting.add(f"N{row}", rule2)
        
        rule3 = CellIsRule(operator='equal', formula=['"Low"'], fill=low_fill)
        ws.conditional_formatting.add(f"N{row}", rule3)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    row = 106
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "DISTRIBUTION SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Total Distributions:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(B5:B104)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Total Recipients:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=SUM(G5:G104)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Avg Engagement Rate:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=AVERAGE(M5:M104)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '0.0"%"'
    
    row += 1
    ws[f"A{row}"] = "High Effectiveness:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(N5:N104,"High")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "Low Effectiveness:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(N5:N104,"Low")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")


# ============================================================================
# SECTION 9: SHEET 7 - FEEDBACK_COLLECTION
# ============================================================================

def create_feedback_collection(ws, styles, validations):
    """Create Feedback_Collection sheet - stakeholder feedback and ratings."""
    
    # Title
    ws.merge_cells("A1:T1")
    ws["A1"] = "STAKEHOLDER FEEDBACK COLLECTION"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:T2")
    ws["A2"] = "Capture stakeholder feedback on intelligence products - ratings, actions taken, improvement suggestions. Use 1-5 scale for all ratings."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Feedback ID",
        "Distribution ID",
        "Product ID",
        "Stakeholder ID",
        "Stakeholder Name",
        "Feedback Date",
        "Feedback Method",
        "Overall Rating",
        "Relevance Rating",
        "Timeliness Rating",
        "Actionability Rating",
        "Clarity Rating",
        "Intelligence Used",
        "Action Taken",
        "Improvement Suggestions",
        "Would Recommend",
        "Positive Comments",
        "Negative Comments",
        "Follow Up Required",
        "Notes",
    ]
    
    col_widths = [15, 15, 15, 15, 25, 18, 18, 15, 18, 18, 20, 15, 18, 30, 30, 18, 30, 30, 18, 30]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Data rows (100 feedback entries capacity)
    for row in range(5, 105):
        # Feedback_ID (auto-generated format)
        ws[f"A{row}"] = f"FB-2025-{row-4:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # Distribution_ID (B)
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Product_ID (C) - can be VLOOKUP from Distribution Tracking
        ws[f"C{row}"] = "=IFERROR(VLOOKUP(B" + str(row) + ",'Distribution Tracking'!A:B,2,FALSE),\"\")"
        ws[f"C{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"C{row}"].border = get_border()

        # Stakeholder_ID (D)
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = get_border()

        # Stakeholder_Name (E) - formula from Stakeholder Registry
        ws[f"E{row}"] = "=IFERROR(VLOOKUP(D" + str(row) + ",'Stakeholder Registry'!A:B,2,FALSE),\"\")"
        ws[f"E{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"E{row}"].border = get_border()
        
        # Feedback_Date (F)
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        ws[f"F{row}"].number_format = 'DD.MM.YYYY'
        
        # Feedback_Method dropdown (G)
        feedback_method_dv = DataValidation(
            type="list",
            formula1='"Survey,Email_Reply,Meeting,Ticket,Call"',
            allow_blank=False
        )
        feedback_method_dv.add(ws[f"G{row}"])
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        
        # Rating fields (H-L) - numbers 1-5
        for col in ["H", "I", "J", "K", "L"]:
            rating_dv = DataValidation(
                type="whole",
                operator="between",
                formula1="1",
                formula2="5",
                allow_blank=True
            )
            rating_dv.add(ws[f"{col}{row}"])
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
            ws[f"{col}{row}"].number_format = '0'
        
        # Intelligence_Used dropdown (M)
        used_dv = DataValidation(
            type="list",
            formula1='"Yes,Partial,No,Will_Use_Later"',
            allow_blank=False
        )
        used_dv.add(ws[f"M{row}"])
        ws[f"M{row}"].fill = styles["input_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        
        # Action_Taken, Improvement_Suggestions (N-O)
        for col in ["N", "O"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
        
        # Would_Recommend dropdown (P)
        recommend_dv = DataValidation(
            type="list",
            formula1='"Yes,Maybe,No"',
            allow_blank=True
        )
        recommend_dv.add(ws[f"P{row}"])
        ws[f"P{row}"].fill = styles["input_cell"]["fill"]
        ws[f"P{row}"].border = get_border()
        
        # Positive_Comments, Negative_Comments (Q-R)
        for col in ["Q", "R"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
        
        # Follow_Up_Required dropdown (S)
        followup_dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        followup_dv.add(ws[f"S{row}"])
        ws[f"S{row}"].fill = styles["input_cell"]["fill"]
        ws[f"S{row}"].border = get_border()
        
        # Notes (T)
        ws[f"T{row}"].fill = styles["input_cell"]["fill"]
        ws[f"T{row}"].border = get_border()
    
    # Apply custom data validations
    for row in range(5, 105):
        feedback_method_dv = DataValidation(
            type="list",
            formula1='"Survey,Email_Reply,Meeting,Ticket,Call"',
            allow_blank=False
        )
        ws.add_data_validation(feedback_method_dv)
        feedback_method_dv.add(ws[f"G{row}"])
        
        used_dv = DataValidation(
            type="list",
            formula1='"Yes,Partial,No,Will_Use_Later"',
            allow_blank=False
        )
        ws.add_data_validation(used_dv)
        used_dv.add(ws[f"M{row}"])
        
        recommend_dv = DataValidation(
            type="list",
            formula1='"Yes,Maybe,No"',
            allow_blank=True
        )
        ws.add_data_validation(recommend_dv)
        recommend_dv.add(ws[f"P{row}"])
        
        followup_dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        ws.add_data_validation(followup_dv)
        followup_dv.add(ws[f"S{row}"])
        
        # Rating validations (1-5)
        for col in ["H", "I", "J", "K", "L"]:
            rating_dv = DataValidation(
                type="whole",
                operator="between",
                formula1="1",
                formula2="5",
                allow_blank=True
            )
            ws.add_data_validation(rating_dv)
            rating_dv.add(ws[f"{col}{row}"])
    
    # Conditional formatting for ratings
    # 5 = Excellent (Green), 4 = Good (Light Green), 3 = Fair (Yellow), 1-2 = Poor (Red)
    excellent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fair_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    poor_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in range(5, 105):
        for col in ["H", "I", "J", "K", "L"]:
            rule1 = CellIsRule(operator='equal', formula=['5'], fill=excellent_fill)
            ws.conditional_formatting.add(f"{col}{row}", rule1)
            
            rule2 = CellIsRule(operator='equal', formula=['4'], fill=good_fill)
            ws.conditional_formatting.add(f"{col}{row}", rule2)
            
            rule3 = CellIsRule(operator='equal', formula=['3'], fill=fair_fill)
            ws.conditional_formatting.add(f"{col}{row}", rule3)
            
            rule4 = CellIsRule(operator='lessThan', formula=['3'], fill=poor_fill)
            ws.conditional_formatting.add(f"{col}{row}", rule4)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    row = 106
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "FEEDBACK SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Total Feedback:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(D5:D104)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Avg Overall Rating:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=AVERAGE(H5:H104)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '0.0'
    
    row += 1
    ws[f"A{row}"] = "Avg Relevance Rating:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=AVERAGE(I5:I104)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '0.0'
    
    row += 1
    ws[f"A{row}"] = "Avg Timeliness Rating:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=AVERAGE(J5:J104)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '0.0'
    
    row += 1
    ws[f"A{row}"] = "Intelligence Used (Yes):"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(M5:M104,"Yes")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "Would Recommend (Yes):"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(P5:P104,"Yes")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")


# ============================================================================
# SECTION 10: SHEET 8 - INTEGRATION_METRICS
# ============================================================================

def create_integration_metrics(ws, styles, validations):
    """Create Integration_Metrics sheet - KPIs for integration effectiveness."""
    
    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "INTEGRATION & DISSEMINATION KPIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:N2")
    ws["A2"] = "Track key performance indicators for tool integration, IOC effectiveness, dissemination reach, and stakeholder engagement."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Metric ID",
        "Metric Name",
        "Metric Category",
        "Measurement Period",
        "Target Value",
        "Actual Value",
        "Unit",
        "Performance vs Target",
        "Status",
        "Trend",
        "Data Source",
        "Owner",
        "Last Updated",
        "Notes",
    ]
    
    col_widths = [12, 35, 22, 20, 15, 15, 10, 22, 18, 15, 30, 25, 15, 35]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Pre-populate standard metrics
    standard_metrics = [
        ("IM-001", "% of Security Tools with TI Integration", "Tool_Integration", "Monthly", 80, 85, "%"),
        ("IM-002", "IOC Deployment Success Rate", "IOC_Effectiveness", "Monthly", 95, 98, "%"),
        ("IM-003", "Avg Time from IOC Publication to Deployment", "IOC_Effectiveness", "Monthly", 4, 2.5, "hours"),
        ("IM-004", "IOC Hit Rate (per 100 IOCs)", "IOC_Effectiveness", "Monthly", 15, 18, "hits"),
        ("IM-005", "False Positive Rate for IOCs", "IOC_Effectiveness", "Monthly", 5, 3, "%"),
        ("IM-006", "Intelligence Distribution Reach", "Dissemination", "Monthly", 90, 92, "%"),
        ("IM-007", "Average Stakeholder Engagement Score", "Stakeholder_Engagement", "Monthly", 4.0, 4.3, "score"),
        ("IM-008", "Feedback Response Rate", "Stakeholder_Engagement", "Monthly", 30, 35, "%"),
        ("IM-009", "Intelligence Consumption Rate (% acted upon)", "Stakeholder_Engagement", "Monthly", 40, 45, "%"),
        ("IM-010", "Mean Time from Intelligence to Action", "Automation", "Monthly", 24, 18, "hours"),
    ]
    
    row = 5
    for metric_id, name, category, period, target, actual, unit in standard_metrics:
        # Metric_ID (A)
        ws[f"A{row}"] = metric_id
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # Metric_Name (B)
        ws[f"B{row}"] = name
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Metric_Category dropdown (C)
        category_dv = DataValidation(
            type="list",
            formula1='"Tool_Integration,IOC_Effectiveness,Dissemination,Stakeholder_Engagement,Automation"',
            allow_blank=False
        )
        category_dv.add(ws[f"C{row}"])
        ws[f"C{row}"] = category
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        # Measurement_Period dropdown (D)
        period_dv = DataValidation(
            type="list",
            formula1='"Daily,Weekly,Monthly,Quarterly"',
            allow_blank=False
        )
        period_dv.add(ws[f"D{row}"])
        ws[f"D{row}"] = period
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        # Target_Value (E)
        ws[f"E{row}"] = target
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = get_border()
        ws[f"E{row}"].number_format = '0.0'
        
        # Actual_Value (F)
        ws[f"F{row}"] = actual
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        ws[f"F{row}"].number_format = '0.0'
        
        # Unit (G)
        ws[f"G{row}"] = unit
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        
        # Performance_vs_Target formula (H)
        ws[f"H{row}"] = f'=IF(AND(E{row}<>"",F{row}<>""),((F{row}/E{row})-1)*100,"")'
        ws[f"H{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        ws[f"H{row}"].number_format = '+0.0%;-0.0%;0.0%'
        
        # Status formula (I)
        ws[f"I{row}"] = f'=IF(H{row}="","",IF(H{row}>=10,"Exceeds Target",IF(H{row}>=0,"Meets Target",IF(H{row}>=-10,"Below Target","Critical"))))'
        ws[f"I{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        
        # Trend dropdown (J)
        trend_dv = DataValidation(
            type="list",
            formula1='"Improving,Stable,Declining"',
            allow_blank=True
        )
        trend_dv.add(ws[f"J{row}"])
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        
        # Data_Source, Owner, Last_Updated, Notes (K-N)
        for col in ["K", "L", "N"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
        
        ws[f"M{row}"].fill = styles["input_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        ws[f"M{row}"].number_format = 'DD.MM.YYYY'
        
        row += 1
    
    # Add 20 more empty rows for custom metrics
    for i in range(row, row + 20):
        ws[f"A{i}"].font = Font(bold=True, size=9)
        ws[f"A{i}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{i}"].border = get_border()
        
        # Setup all dropdowns and formulas for empty rows
        for col in ["B", "E", "F", "G", "K", "L", "N"]:
            ws[f"{col}{i}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{i}"].border = get_border()
        
        ws[f"M{i}"].fill = styles["input_cell"]["fill"]
        ws[f"M{i}"].border = get_border()
        ws[f"M{i}"].number_format = 'DD.MM.YYYY'
        
        # Formulas
        ws[f"H{i}"] = f'=IF(AND(E{i}<>"",F{i}<>""),((F{i}/E{i})-1)*100,"")'
        ws[f"H{i}"].fill = styles["formula_cell"]["fill"]
        ws[f"H{i}"].border = get_border()
        ws[f"H{i}"].number_format = '+0.0%;-0.0%;0.0%'
        
        ws[f"I{i}"] = f'=IF(H{i}="","",IF(H{i}>=10,"Exceeds Target",IF(H{i}>=0,"Meets Target",IF(H{i}>=-10,"Below Target","Critical"))))'
        ws[f"I{i}"].fill = styles["formula_cell"]["fill"]
        ws[f"I{i}"].border = get_border()
    
    # Apply custom data validations for all rows
    for i in range(5, row + 20):
        category_dv = DataValidation(
            type="list",
            formula1='"Tool_Integration,IOC_Effectiveness,Dissemination,Stakeholder_Engagement,Automation"',
            allow_blank=False
        )
        ws.add_data_validation(category_dv)
        category_dv.add(ws[f"C{i}"])
        
        period_dv = DataValidation(
            type="list",
            formula1='"Daily,Weekly,Monthly,Quarterly"',
            allow_blank=False
        )
        ws.add_data_validation(period_dv)
        period_dv.add(ws[f"D{i}"])
        
        trend_dv = DataValidation(
            type="list",
            formula1='"Improving,Stable,Declining"',
            allow_blank=True
        )
        ws.add_data_validation(trend_dv)
        trend_dv.add(ws[f"J{i}"])
    
    # Conditional formatting for Status
    exceeds_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    meets_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    below_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    critical_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for i in range(5, row + 20):
        rule1 = CellIsRule(operator='equal', formula=['"Exceeds Target"'], fill=exceeds_fill)
        ws.conditional_formatting.add(f"I{i}", rule1)

        rule2 = CellIsRule(operator='equal', formula=['"Meets Target"'], fill=meets_fill)
        ws.conditional_formatting.add(f"I{i}", rule2)

        rule3 = CellIsRule(operator='equal', formula=['"Below Target"'], fill=below_fill)
        ws.conditional_formatting.add(f"I{i}", rule3)
        
        rule4 = CellIsRule(operator='equal', formula=['"Critical"'], fill=critical_fill)
        ws.conditional_formatting.add(f"I{i}", rule4)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    summary_row = row + 21
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    ws[f"A{summary_row}"] = "METRICS SUMMARY"
    ws[f"A{summary_row}"].font = styles["section_header"]["font"]
    ws[f"A{summary_row}"].fill = styles["section_header"]["fill"]
    
    summary_row += 1
    ws[f"A{summary_row}"] = "Exceeds Target:"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"] = f'=COUNTIF(I5:I{row+19},"Exceeds Target")'
    ws[f"B{summary_row}"].font = Font(bold=True, color="006100")

    summary_row += 1
    ws[f"A{summary_row}"] = "Meets Target:"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"] = f'=COUNTIF(I5:I{row+19},"Meets Target")'
    ws[f"B{summary_row}"].font = Font(bold=True, color="006100")

    summary_row += 1
    ws[f"A{summary_row}"] = "Below Target:"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"] = f'=COUNTIF(I5:I{row+19},"Below Target")'
    ws[f"B{summary_row}"].font = Font(bold=True, color="C65911")
    
    summary_row += 1
    ws[f"A{summary_row}"] = "Critical:"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"] = f'=COUNTIF(I5:I{row+19},"Critical")'
    ws[f"B{summary_row}"].font = Font(bold=True, color="9C0006")

# ============================================================================
# SECTION 11: SHEET 9 - ACTION_ITEMS
# ============================================================================

def create_action_items(ws, styles, validations):
    """Create Action_Items sheet - integration improvement tasks."""
    
    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "INTEGRATION & DISTRIBUTION ACTION ITEMS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:N2")
    ws["A2"] = "Track remediation tasks, improvements, and follow-up actions for tool integration, IOC effectiveness, and stakeholder engagement."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Action ID",
        "Issue Type",
        "Issue Description",
        "Priority",
        "Assigned To",
        "Due Date",
        "Status",
        "Status Notes",
        "Resolution Date",
        "Evidence Link",
        "Created Date",
        "Created By",
        "Last Updated",
        "Notes",
    ]
    
    col_widths = [15, 20, 35, 12, 25, 15, 15, 30, 15, 30, 15, 25, 15, 30]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Data rows (50 action items capacity)
    for row in range(5, 55):
        # Action_ID (auto-generated format)
        ws[f"A{row}"] = f"ACT-2025-{row-4:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"A{row}"].border = get_border()
        
        # Issue_Type dropdown (B)
        issue_type_dv = DataValidation(
            type="list",
            formula1='"Tool_Integration,IOC_Quality,Dissemination,Stakeholder_Engagement,Process,Automation,Other"',
            allow_blank=False
        )
        issue_type_dv.add(ws[f"B{row}"])
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Issue_Description (C)
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        # Priority dropdown (D)
        validations['priority'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        # Assigned_To (E)
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = get_border()
        
        # Due_Date (F)
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        ws[f"F{row}"].number_format = 'DD.MM.YYYY'
        
        # Status dropdown (G)
        validations['action_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        
        # Status_Notes (H)
        ws[f"H{row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        
        # Resolution_Date (I)
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        ws[f"I{row}"].number_format = 'DD.MM.YYYY'
        
        # Evidence_Link (J)
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        
        # Created_Date (K)
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()
        ws[f"K{row}"].number_format = 'DD.MM.YYYY'
        
        # Created_By (L)
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()
        
        # Last_Updated (M)
        ws[f"M{row}"].fill = styles["input_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        ws[f"M{row}"].number_format = 'DD.MM.YYYY'
        
        # Notes (N)
        ws[f"N{row}"].fill = styles["input_cell"]["fill"]
        ws[f"N{row}"].border = get_border()
    
    # Apply custom data validations
    for row in range(5, 55):
        issue_type_dv = DataValidation(
            type="list",
            formula1='"Tool_Integration,IOC_Quality,Dissemination,Stakeholder_Engagement,Process,Automation,Other"',
            allow_blank=False
        )
        ws.add_data_validation(issue_type_dv)
        issue_type_dv.add(ws[f"B{row}"])
    
    # Apply standard validations
    ws.add_data_validation(validations['priority'])
    ws.add_data_validation(validations['action_status'])
    
    # Conditional formatting for Priority
    critical_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    high_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    for row in range(5, 55):
        rule1 = CellIsRule(operator='equal', formula=['"Critical"'], fill=critical_fill)
        ws.conditional_formatting.add(f"D{row}", rule1)
        
        rule2 = CellIsRule(operator='equal', formula=['"High"'], fill=high_fill)
        ws.conditional_formatting.add(f"D{row}", rule2)
        
        rule3 = CellIsRule(operator='equal', formula=['"Medium"'], fill=medium_fill)
        ws.conditional_formatting.add(f"D{row}", rule3)
        
        rule4 = CellIsRule(operator='equal', formula=['"Low"'], fill=low_fill)
        ws.conditional_formatting.add(f"D{row}", rule4)
    
    # Status conditional formatting
    resolved_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    blocked_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    for row in range(5, 55):
        rule1 = CellIsRule(operator='equal', formula=['"Resolved"'], fill=resolved_fill)
        ws.conditional_formatting.add(f"G{row}", rule1)
        
        rule2 = CellIsRule(operator='equal', formula=['"Closed"'], fill=resolved_fill)
        ws.conditional_formatting.add(f"G{row}", rule2)
        
        rule3 = CellIsRule(operator='equal', formula=['"Blocked"'], fill=blocked_fill)
        ws.conditional_formatting.add(f"G{row}", rule3)
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Summary dashboard
    row = 56
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ACTION ITEMS SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Total Actions:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(C5:C54)'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Critical Priority:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(D5:D54,"Critical")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")
    
    row += 1
    ws[f"A{row}"] = "High Priority:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(D5:D54,"High")'
    ws[f"B{row}"].font = Font(bold=True, color="C65911")
    
    row += 1
    ws[f"A{row}"] = "Open Actions:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"Open")'
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "In Progress:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"In_Progress")'
    ws[f"B{row}"].font = Font(bold=True, color="0066CC")
    
    row += 1
    ws[f"A{row}"] = "Blocked:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"Blocked")'
    ws[f"B{row}"].font = Font(bold=True, color="C65911")
    
    row += 1
    ws[f"A{row}"] = "Resolved/Closed:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(G5:G54,"Resolved")+COUNTIF(G5:G54,"Closed")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")
    
    row += 1
    ws[f"A{row}"] = "Overdue Actions:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIFS(F5:F54,"<"&TODAY(),G5:G54,"<>Resolved",G5:G54,"<>Closed")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")


# ============================================================================
# SECTION 12: SHEET 10 - METADATA
# ============================================================================

def create_metadata(ws, styles):
    """Create Metadata sheet - workbook generation and version tracking."""
    
    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = "WORKBOOK METADATA"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    row = 3
    
    # Document Information
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:B{row}")
    
    row += 1
    metadata_info = [
        ("Document ID", "ISMS-IMP-A.5.7.3"),
        ("Title", "Intelligence Integration & Distribution Assessment"),
        ("Version", "1.0"),
        ("Related Policy", "ISMS-POL-A.5.7 (Threat Intelligence Policy)"),
        ("ISO Control", "ISO/IEC 27001:2022 Annex A Control 5.7"),
        ("Review Cycle", "Quarterly"),
    ]
    
    for label, value in metadata_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1
    
    row += 1
    
    # Generation Information
    ws[f"A{row}"] = "GENERATION INFORMATION"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:B{row}")
    
    row += 1
    generation_info = [
        ("Generation Date", datetime.now().strftime("%d.%m.%Y %H:%M:%S")),
        ("Generator Script", "generate_a57_3_integration.py"),
        ("Script Version", "1.0"),
        ("Generated By", "Python openpyxl"),
        ("Python Version", "3.9+"),
    ]
    
    for label, value in generation_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1
    
    row += 1
    
    # Workbook Structure
    ws[f"A{row}"] = "WORKBOOK STRUCTURE"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:B{row}")
    
    row += 1
    ws[f"A{row}"] = "Sheet"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "Purpose"
    ws[f"B{row}"].font = Font(bold=True)
    
    row += 1
    sheets_info = [
        ("1. Instructions", "Assessment guidance and TLP reference"),
        ("2. Tool_Integration_Matrix", "Security tool integration status (30 tools)"),
        ("3. IOC_Deployment", "IOC deployment and effectiveness tracking (100 IOCs)"),
        ("4. Dissemination_Channels", "Intelligence distribution mechanisms (25 channels)"),
        ("5. Stakeholder_Registry", "Audience management (50 stakeholders)"),
        ("6. Distribution_Tracking", "Intelligence distribution records (100 distributions)"),
        ("7. Feedback_Collection", "Stakeholder feedback and ratings (100 entries)"),
        ("8. Integration_Metrics", "KPIs for integration effectiveness (30 metrics)"),
        ("9. Action_Items", "Integration improvement tasks (50 actions)"),
        ("10. Metadata", "This sheet - generation and version information"),
    ]
    
    for sheet, purpose in sheets_info:
        ws[f"A{row}"] = sheet
        ws[f"B{row}"] = purpose
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 1
    
    # Integration Points
    ws[f"A{row}"] = "INTEGRATION POINTS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:B{row}")
    
    row += 1
    integration_info = [
        ("From ISMS-IMP-A.5.7.2", "Intelligence production feeds distribution tracking"),
        ("To ISMS-IMP-A.5.7.1", "IOC effectiveness feeds source quality evaluation"),
        ("To Control 8.8", "Vulnerability-related IOCs and tool integration status"),
    ]
    
    for integration, description in integration_info:
        ws[f"A{row}"] = integration
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = description
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 2
    
    # Footer note
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "This workbook was auto-generated. Manual modifications will be lost on regeneration. Update source data and regenerate instead."
    ws[f"A{row}"].font = Font(italic=True, size=9, color="9C0006")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60


# ============================================================================
# SECTION 13: MAIN GENERATION FUNCTION
# ============================================================================



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Tool Integration Matrix — map threat intelligence feeds to security tools (SIEM, EDR, firewall).',
        '2. Complete IOC Deployment — track indicators of compromise pushed to detection systems.',
        '3. Document Dissemination Channels — list all channels used to share intelligence internally.',
        '4. Complete Stakeholder Registry — identify all consumers of threat intelligence outputs.',
        '5. Track Distribution in Distribution Tracking — confirm intelligence reaches all relevant teams.',
        '6. Complete Feedback Collection — record stakeholder feedback on intelligence utility.',
        '7. Review Integration Metrics — measure effectiveness of threat intelligence integration.',
        '8. Maintain the Evidence Register with integration records and feedback documentation.',
        '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A23"] = "Status Legend"
    ws["A23"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=24, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 25 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(ws, styles):
    """Summary Dashboard — A.5.7.3 Threat Intelligence Integration & Distribution.

    Gold Standard implementation per A.8.33-34 reference pattern:
    - A1: em dash title with "— SUMMARY DASHBOARD" (GS-SD-014)
    - A2: 003366 italic, left-aligned, NO fill
    - TABLE 1: 003366 banner, 4472C4 section headers, D9D9D9 col headers, FFFFCC data
    - TABLE 2: 003366 banner, D9D9D9 col headers (NOT 003366), FFFFFF data rows
    - TABLE 3: C00000 banner, D9D9D9 col headers, FFFFFF data rows
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    def _f(h):
        return PatternFill("solid", fgColor=h)

    def _b():
        t = Side(style="thin")
        return Border(left=t, right=t, top=t, bottom=t)

    def _banner(r, v, fill, fc="FFFFFF", sz=12):
        """Render a full-width section banner (cols A:F merged)."""
        cell = ws.cell(row=r, column=1)
        cell.value = v
        cell.font = Font(name="Calibri", bold=True, color=fc, size=sz)
        cell.fill = _f(fill)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{r}:F{r}")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _b()

    def _col_hdr(r, labels):
        """Render D9D9D9 column headers."""
        for i, lbl in enumerate(labels):
            cell = ws.cell(row=r, column=1 + i)
            cell.value = lbl
            cell.font = Font(name="Calibri", bold=True, color="000000", size=10)
            cell.fill = _f("D9D9D9")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _b()

    def _dat(r, c, v, fill="FFFFCC", fc="000000", bold=False, num=False):
        cell = ws.cell(row=r, column=c)
        cell.value = v
        cell.font = Font(name="Calibri", bold=bold, color=fc)
        cell.fill = _f(fill)
        cell.alignment = Alignment(
            horizontal="center" if num else "left",
            vertical="center", wrap_text=True
        )
        cell.border = _b()
        return cell

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A3"

    # ── Row 1: Title (GS-SD-014: must contain "— SUMMARY DASHBOARD") ────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ISMS-IMP-A.5.7.3 \u2014 THREAT INTELLIGENCE INTEGRATION & DISTRIBUTION \u2014 SUMMARY DASHBOARD"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.fill = _f("003366")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, 7):
        ws.cell(row=1, column=col).border = _b()
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (003366 italic, left-aligned, NO fill) ──────────────
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "ISO/IEC 27001:2022 | Control A.5.7 | Tool integration status, IOC deployment health, and action tracking"
    c.font = Font(name="Calibri", italic=True, color="003366", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 1 — STATUS DISTRIBUTION
    # ══════════════════════════════════════════════════════════════════════
    _banner(4, "TABLE 1: STATUS DISTRIBUTION", "003366")

    # Section A — Tool Integration Status (col G, rows 5:34)
    _banner(5, "Section A: Tool Integration Status ('Tool Integration Matrix' \u2014 Column G, rows 5:34)", "4472C4", sz=10)
    _col_hdr(6, ["Status", "Count", "% of Total"])

    tool_statuses = ["Fully_Integrated", "Partially_Integrated", "Planned", "Not_Integrated"]
    for i, status in enumerate(tool_statuses):
        r = 7 + i
        _dat(r, 1, status)
        _dat(r, 2, f'=COUNTIF(\'Tool Integration Matrix\'!G5:G34,\"{status}\")', num=True)
        _dat(r, 3, f'=IF($B$11=0,"\u2014",TEXT(B{r}/$B$11,"0.0%"))', num=True)

    # TOTAL (row 11)
    _dat(11, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(11, 2, "=SUM(B7:B10)", "D9D9D9", num=True)
    _dat(11, 3, "\u2014", "D9D9D9", num=True)

    # Section B — IOC Deployment Status (col N, rows 5:104)
    _banner(12, "Section B: IOC Deployment Status ('IOC Deployment' \u2014 Column N, rows 5:104)", "4472C4", sz=10)
    _col_hdr(13, ["Status", "Count", "% of Total"])

    ioc_statuses = ["Active", "Expired", "Withdrawn", "Under_Review"]
    for i, status in enumerate(ioc_statuses):
        r = 14 + i
        _dat(r, 1, status)
        _dat(r, 2, f'=COUNTIF(\'IOC Deployment\'!N5:N104,\"{status}\")', num=True)
        _dat(r, 3, f'=IF($B$18=0,"\u2014",TEXT(B{r}/$B$18,"0.0%"))', num=True)

    # TOTAL (row 18)
    _dat(18, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(18, 2, "=SUM(B14:B17)", "D9D9D9", num=True)
    _dat(18, 3, "\u2014", "D9D9D9", num=True)

    # Row 19: buffer
    for c in range(1, 7):
        ws.cell(row=19, column=c).fill = _f("FFFFFF")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 2 — KEY PERFORMANCE INDICATORS
    # Gold Standard: 003366 banner, D9D9D9 col headers, FFFFFF data rows
    # ══════════════════════════════════════════════════════════════════════
    _banner(20, "TABLE 2: KEY PERFORMANCE INDICATORS", "003366")
    _col_hdr(21, ["KPI Metric", "Value", "Notes"])
    ws.merge_cells("C21:F21")
    ws.cell(row=21, column=3).alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("Total Tools in Scope",
         "=COUNTA('Tool Integration Matrix'!A5:A34)",
         "All security tools tracked in integration matrix"),
        ("Fully Integrated Tools",
         "=COUNTIF('Tool Integration Matrix'!G5:G34,\"Fully_Integrated\")",
         "Tools with complete threat intelligence integration"),
        ("Not Integrated Tools",
         "=COUNTIF('Tool Integration Matrix'!G5:G34,\"Not_Integrated\")",
         "Tools with no integration \u2014 gap requiring remediation"),
        ("Active IOC Deployments",
         "=COUNTIF('IOC Deployment'!N5:N104,\"Active\")",
         "IOCs currently deployed and active"),
        ("Open Action Items",
         "=COUNTIF('Action Items'!G5:G54,\"Open\")",
         "Actions not yet started \u2014 requires attention"),
        ("Closed Action Items",
         "=COUNTIF('Action Items'!G5:G54,\"Closed\")",
         "Successfully closed improvement actions"),
    ]
    for i, (metric, formula, note) in enumerate(kpis):
        r = 22 + i
        # Col A: metric label — FFFFFF fill, 000000 font, NOT bold (GS-SD-015)
        cell_a = ws.cell(row=r, column=1)
        cell_a.value = metric
        cell_a.font = Font(name="Calibri", color="000000")
        cell_a.fill = _f("FFFFFF")
        cell_a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_a.border = _b()
        # Col B: value formula — FFFFFF fill
        cell_b = ws.cell(row=r, column=2)
        cell_b.value = formula
        cell_b.font = Font(name="Calibri", color="000000")
        cell_b.fill = _f("FFFFFF")
        cell_b.alignment = Alignment(horizontal="center", vertical="center")
        cell_b.border = _b()
        # Col C:F merged: notes — FFFFCC fill
        cell_c = ws.cell(row=r, column=3)
        cell_c.value = note
        cell_c.font = Font(name="Calibri", color="000000")
        cell_c.fill = _f("FFFFFF")
        cell_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_c.border = _b()
        ws.merge_cells(f"C{r}:F{r}")

    # Row 28: buffer
    for c in range(1, 7):
        ws.cell(row=28, column=c).fill = _f("FFFFFF")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 3 — CRITICAL FINDINGS
    # Gold Standard: C00000 banner, D9D9D9 col headers, FFFFFF data rows
    # ══════════════════════════════════════════════════════════════════════
    _banner(29, "TABLE 3: CRITICAL FINDINGS", "C00000")
    _col_hdr(30, ["Finding", "Count", "Action Required"])
    ws.merge_cells("C30:F30")
    ws.cell(row=30, column=3).alignment = Alignment(horizontal="center", vertical="center")

    critical = [
        ("Not Integrated Tools",
         "=COUNTIF('Tool Integration Matrix'!G5:G34,\"Not_Integrated\")",
         "Define integration plan and schedule for each unintegrated tool"),
        ("Expired or Withdrawn IOCs",
         "=COUNTIF('IOC Deployment'!N5:N104,\"Expired\")+COUNTIF('IOC Deployment'!N5:N104,\"Withdrawn\")",
         "Remove expired/withdrawn IOCs from active feeds and update records"),
        ("Open Action Items",
         "=COUNTIF('Action Items'!G5:G54,\"Open\")",
         "Review and assign all open actions; set realistic due dates"),
    ]
    for i, (finding, formula, action) in enumerate(critical):
        r = 31 + i
        c1 = ws.cell(row=r, column=1)
        c1.value = finding
        c1.font = Font(name="Calibri", bold=True, color="000000")
        c1.fill = _f("FFFFCC")
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c1.border = _b()

        c2 = ws.cell(row=r, column=2)
        c2.value = formula
        c2.font = Font(name="Calibri", color="000000")
        c2.fill = _f("FFFFCC")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _b()

        c3 = ws.cell(row=r, column=3)
        c3.value = action
        c3.font = Font(name="Calibri", color="000000")
        c3.fill = _f("FFFFCC")
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3.border = _b()
        ws.merge_cells(f"C{r}:F{r}")


def create_evidence_register(ws):
    """Create Evidence Register worksheet (Gold Standard — 100 data rows, navy headers)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border

    # Row 2: SUBTITLE
    ws.merge_cells('A2:H2')
    ws['A2'] = f'{DOCUMENT_ID} | {CONTROL_REF}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    # Row 3: intentionally empty (visual separator)

    # Row 4: COLUMN HEADERS with 003366 fill
    headers = [
        'Evidence ID', 'Assessment Area', 'Evidence Type', 'Description',
        'Location / Path', 'Date Collected', 'Collected By', 'Verification Status',
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Data validations
    ev_type_dv = DataValidation(
        type='list',
        formula1='"Config File,Screenshot,Report,Log File,Test Result,Policy Document,Contract,Diagram,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type='list',
        formula1='"\u2705 Verified,Pending,\u274c Not Verified"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # Row 5: SAMPLE ROW (F2F2F2 grey)
    sample_data = {
        1: 'EV-001',
        2: 'Tool Integration Matrix',
        3: 'Screenshot',
        4: 'SIEM integration configuration showing STIX/TAXII feed subscription and IOC ingestion',
        5: '/evidence/a57/siem-integration-config-2026-01-15.png',
        6: datetime.now().strftime('%d.%m.%Y'),
        7: 'Security Manager',
        8: '\u2705 Verified',
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.border = border
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(
            horizontal='center' if col == 1 else 'left',
            vertical='center',
            wrap_text=True
        )
    ev_type_dv.add(ws['C5'])
    ver_status_dv.add(ws['H5'])

    # Rows 6-105: EMPTY DATA ROWS (FFFFCC yellow, 100 rows)
    for data_row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(
                horizontal='center' if col == 1 else 'left',
                vertical='center',
                wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f'C{data_row}'])
        ver_status_dv.add(ws[f'H{data_row}'])

    # Column widths and freeze
    for col, width in [('A', 15), ('B', 28), ('C', 22), ('D', 45), ('E', 45), ('F', 16), ('G', 22), ('H', 20)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A5'


def create_approval_sheet(ws):
    """Create Approval Sign-Off worksheet (Gold Standard)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL AND SIGN-OFF'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2: CONTROL REFERENCE subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f'{DOCUMENT_ID} | {CONTROL_REF}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells('A3:E3')
    ws['A3'] = 'ASSESSMENT SUMMARY'
    ws['A3'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # Summary fields
    summary_fields = [
        ('Document:', f'{DOCUMENT_ID} - {WORKBOOK_NAME}'),
        ('Assessment Period:', ''),
        ('Overall Compliance Rate:', ''),
        ('Assessment Status:', ''),
        ('Assessed By:', ''),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == '')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        if 'Assessment Status' in label:
            status_row_for_dv = row
        row += 1

    # Assessment Status dropdown
    status_dv = DataValidation(
        type='list',
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f'B{status_row_for_dv}')

    # B6: Overall Compliance Rate — Fully Integrated tools / Total tools
    ws['B6'] = "=IFERROR('Summary Dashboard'!B7/'Summary Dashboard'!B11,\"\")"
    ws['B6'].number_format = '0.0%'

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ['Name:', 'Title:', 'Date:', 'Signature:', 'Comments:']:
            ws[f'A{current_row}'] = field
            ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'B{current_row}:E{current_row}')
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, 'COMPLETED BY (ASSESSOR)', '4472C4')
    row = _create_approver_section(row, 'REVIEWED BY (INFORMATION SECURITY OFFICER)', '4472C4')
    row = _create_approver_section(row, 'APPROVED BY (CISO)', '003366')

    # FINAL DECISION
    ws[f'A{row}'] = 'FINAL DECISION:'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border
    ws.merge_cells(f'B{row}:E{row}')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type='list',
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ['Next Review Date:', 'Review Responsible:', 'Special Considerations:']:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths and freeze
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.freeze_panes = 'A3'


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def main():
    """
    Main function to generate the ISMS-IMP-A.5.7.3 workbook.
    
    Returns:
        str: Path to the generated workbook
    """
    logger.info("")
    logger.info("=" * 70)
    logger.info("ISMS-IMP-A.5.7.3 - Integration & Distribution Assessment Generator")
    logger.info("ISO/IEC 27001:2022 - Control A.5.7: Threat Intelligence")
    logger.info("=" * 70)
    logger.info("")
    
    # Create workbook and setup
    logger.info("[1/14] Creating workbook structure...")
    wb = create_workbook()
    
    logger.info("[2/14] Setting up styles...")
    styles = _STYLES
    
    logger.info("[3/14] Configuring data validations...")
    validations = setup_validations()
    
    # Generate sheets
    logger.info("[4/14] Creating Instructions sheet...")
    create_instructions(wb["Instructions & Legend"], styles)
    
    logger.info("[5/14] Creating Tool_Integration_Matrix sheet...")
    create_tool_integration(wb["Tool Integration Matrix"], styles, validations)
    
    logger.info("[6/14] Creating IOC_Deployment sheet...")
    create_ioc_deployment(wb["IOC Deployment"], styles, validations)
    
    logger.info("[7/14] Creating Dissemination_Channels sheet...")
    create_dissemination_channels(wb["Dissemination Channels"], styles, validations)
    
    logger.info("[8/14] Creating Stakeholder_Registry sheet...")
    create_stakeholder_registry(wb["Stakeholder Registry"], styles, validations)
    
    logger.info("[9/14] Creating Distribution_Tracking sheet...")
    create_distribution_tracking(wb["Distribution Tracking"], styles, validations)
    
    logger.info("[10/14] Creating Feedback_Collection sheet...")
    create_feedback_collection(wb["Feedback Collection"], styles, validations)
    
    logger.info("[11/14] Creating Integration_Metrics sheet...")
    create_integration_metrics(wb["Integration Metrics"], styles, validations)
    
    logger.info("[12/14] Creating Action_Items sheet...")
    create_action_items(wb["Action Items"], styles, validations)
    
    logger.info("[10/12] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"])

    logger.info("[11/12] Creating Summary Dashboard sheet...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[12/12] Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb["Approval Sign-Off"])

    logger.info("[12/12] Finalizing workbook...")
    
    # Generate filename with date
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.5.7.3_Integration_{date_str}.xlsx"
    
    # Finalise validations
    finalize_validations(wb)

    # Save workbook
    logger.info("")
    output_path = _wkbk_dir / OUTPUT_FILENAME
    logger.info(f"Saving workbook as: {output_path.name}")
    wb.save(output_path)
    # Print summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("GENERATION COMPLETE")
    logger.info("=" * 70)
    logger.info("")
    logger.info(f"Output file: {output_path.name}")
    logger.info(f"Total sheets: {len(wb.sheetnames)}")
    logger.info("")
    logger.info("Sheet Summary:")
    logger.info("-" * 40)
    for i, sheet in enumerate(wb.sheetnames, 1):
        logger.info(f"  {i:2}. {sheet}")
    logger.info("")
    logger.info("Key Capacities:")
    logger.info("-" * 40)
    logger.info("  • Security Tools: 30 tools")
    logger.info("  • IOC Deployment: 100 IOCs")
    logger.info("  • Dissemination Channels: 25 channels")
    logger.info("  • Stakeholders: 50 stakeholders")
    logger.info("  • Distribution Records: 100 distributions")
    logger.info("  • Feedback Entries: 100 entries")
    logger.info("  • Integration Metrics: 30 KPIs")
    logger.info("  • Action Items: 50 tasks")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("-" * 40)
    logger.info("  1. Open the workbook in Excel/LibreOffice")
    logger.info("  2. Document all Tool_Integration_Matrix entries")
    logger.info("  3. Load IOC_Deployment data from past 90 days")
    logger.info("  4. Configure Dissemination_Channels and stakeholder registry")
    logger.info("  5. Review Distribution_Tracking for intelligence reach")
    logger.info("  6. Collect and analyse Feedback from stakeholders")
    logger.info("  7. Establish baseline Integration_Metrics")
    logger.info("  8. Generate Action_Items for gaps and improvements")
    logger.info("")
    logger.info("Remember: Evidence > Theater")
    logger.info("         'We integrate threat intelligence' means nothing without")
    logger.info("         IOC hit rates, stakeholder engagement metrics, and feedback.")
    logger.info("")
    return output_path


def validate_workbook(filename):
    """
    Basic validation of generated workbook.
    
    Args:
        filename: Path to the workbook to validate
        
    Returns:
        bool: True if validation passes
    """
    from openpyxl import load_workbook
    
    logger.info("")
    logger.info("Running basic validation...")
    logger.info("-" * 40)
    
    try:
        wb = load_workbook(_wkbk_dir / filename)
        
        expected_sheets = [
            "Instructions & Legend",
            "Tool Integration Matrix",
            "IOC Deployment",
            "Dissemination Channels",
            "Stakeholder Registry",
            "Distribution Tracking",
            "Feedback Collection",
            "Integration Metrics",
            "Action Items",
            "Evidence Register",
            "Summary Dashboard",
            "Approval Sign-Off",
        ]

        # Check sheet count (12 sheets)
        if len(wb.sheetnames) != 12:
            logger.error(f"  ✗ Expected 11 sheets, found {len(wb.sheetnames)}")
            return False
        logger.info(f"  ✓ Sheet count: {len(wb.sheetnames)}")
        
        # Check sheet names
        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                logger.info(f"  ✓ Found: {sheet}")
            else:
                logger.error(f"  ✗ Missing: {sheet}")
                return False
        
        # Check Tool_Integration_Matrix has 30 rows
        ti_sheet = wb["Tool Integration Matrix"]
        ti_sheet.sheet_view.showGridLines = False
        ti_count = 0
        for row in range(5, 35):
            if ti_sheet[f"A{row}"].value:
                ti_count += 1
        
        if ti_count == 30:
            logger.info(f"  ✓ Tool_Integration_Matrix: {ti_count} rows")
        else:
            logger.info(f"  ⚠  Tool_Integration_Matrix: {ti_count} rows (expected 30)")
        
        # Check IOC_Deployment has 100 rows
        ioc_sheet = wb["IOC Deployment"]
        ioc_sheet.sheet_view.showGridLines = False
        ioc_count = 0
        for row in range(5, 105):
            if ioc_sheet[f"A{row}"].value:
                ioc_count += 1
        
        if ioc_count == 100:
            logger.info(f"  ✓ IOC_Deployment: {ioc_count} rows")
        else:
            logger.info(f"  ⚠  IOC_Deployment: {ioc_count} rows (expected 100)")
        
        # Check Integration_Metrics has pre-populated metrics
        metrics_sheet = wb["Integration Metrics"]
        metrics_sheet.sheet_view.showGridLines = False
        metrics_count = 0
        for row in range(5, 35):
            if metrics_sheet[f"B{row}"].value:
                metrics_count += 1
        
        if metrics_count >= 10:
            logger.info(f"  ✓ Integration_Metrics: {metrics_count} metrics (10 pre-populated)")
        else:
            logger.info(f"  ⚠  Integration_Metrics: {metrics_count} metrics")
        
        logger.info("")
        logger.info("Validation Result: ✓ PASSED")
        logger.info("")
        
        wb.close()
        return True
        
    except Exception as e:
        logger.error(f"  ✗ Validation error: {str(e)}")
        return False


# ============================================================================
# SECTION 14: ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    """
    Entry point for script execution.

    Usage:
        python3 generate_a57_3_integration.py

    This will generate the assessment workbook in the current directory.
    """
    logger.info("")
    logger.info("=" * 68)
    logger.info("  ISMS Control A.5.7 - Threat Intelligence")
    logger.info("  Workbook 3: Integration & Distribution Assessment")
    logger.info("")
    logger.info("  'Intelligence without integration is just expensive news.'")
    logger.info("                                   - Security Operations Axiom")
    logger.info("=" * 68)
    logger.info("")
    
    try:
        # Generate the workbook
        output_file = main()
        
        # Validate the output
        if validate_workbook(output_file):
            logger.info(f"✓ Successfully generated: {output_file}")
            sys.exit(0)
        else:
            logger.info("⚠  Validation warnings - please review the output")
            sys.exit(1)
            
    except ImportError as e:
        logger.info("")
        logger.error("ERROR: Missing required library")
        logger.info("-" * 40)
        logger.info(f"  {str(e)}")
        logger.info("")
        logger.info("Please install openpyxl:")
        logger.info("  pip install openpyxl")
        logger.info("  or")
        logger.info("  sudo apt install python3-openpyxl")
        logger.info("")
        sys.exit(1)
        
    except Exception as e:
        logger.info("")
        logger.error("ERROR: Generation failed")
        logger.info("-" * 40)
        logger.info(f"  {str(e)}")
        logger.info("")
        import traceback
        traceback.print_exc()
        sys.exit(1)


# =============================================================================
# END OF GENERATOR SCRIPT
# =============================================================================
#
# Document Control:
#   Version: 1.0
#   Created: 05.01.226
#
# Change History:
#   1.0 - Initial version
#       - 10 sheets for integration & distribution assessment
#       - 30-tool integration matrix
#       - 100-IOC deployment tracking
#       - 25 dissemination channels
#       - 50-stakeholder registry
#       - 100 distribution records
#       - 100 feedback entries
#       - 30 integration KPIs (10 pre-populated)
#       - 50-item action tracker
#       - Comprehensive dropdown validations
#       - Conditional formatting for status/priority/ratings
#
# Dependencies:
#   - Python 3.9+
#   - openpyxl >= 3.0.0
#
# Output:
#   ISMS-IMP-A.5.7.3_Integration_YYYYMMDD.xlsx
#
# Integration Points:
#   - From ISMS-IMP-A.5.7.2 (Collection & Analysis)
#   - To ISMS-IMP-A.5.7.1 (Sources - IOC effectiveness feedback)
#   - To Control 8.8 (Vulnerability Management)
#
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
