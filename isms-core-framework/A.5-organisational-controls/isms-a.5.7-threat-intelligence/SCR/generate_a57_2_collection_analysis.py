#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.7.2 - Intelligence Collection & Analysis Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.7: Threat Intelligence
Assessment Domain 2 of 5: Intelligence Collection & Analysis Workflows

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific threat intelligence workflows, analysis processes,
and operational requirements.

Key customization areas:
1. Intelligence collection workflows and automation (match your SIEM/SOAR tools)
2. Analysis team structure and responsibilities (adapt to your org chart)
3. MITRE ATT&CK coverage requirements (align with your threat model)
4. Indicator processing volumes and SLAs (based on operational capacity)
5. Quality assurance thresholds (confidence levels, validation requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.7 Threat Intelligence Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
threat intelligence collection, processing, analysis, and quality assurance
workflows against ISO 27001:2022 Control A.5.7 requirements.

**Purpose:**
Enables systematic assessment of how threat intelligence is collected from
sources (assessed in A.5.7.1), analysed by security teams, enriched with
context, and prepared for operational use in security workflows.

**Assessment Scope:**
- Automated intelligence collection mechanisms (API integrations, feeds)
- Manual intelligence collection workflows (OSINT, vendor portals)
- Indicator processing and normalisation (IOCs, TTPs, CVSS enrichment)
- Threat analysis workflows and analyst capabilities
- MITRE ATT&CK framework mapping and coverage
- Intelligence enrichment and contextualization
- Quality assurance and validation processes
- Analyst productivity and workload management
- Intelligence aging and archival policies
- Integration with SIEM, SOAR, and security tools
- CVSS validation and accuracy improvement (VTL support)
- Audit evidence collection and metrics

**Generated Workbook Structure (12 Sheets):**
1. Instructions - Assessment guidance and workflow methodology
2. Collection_Workflows - Automated and manual collection processes
3. Processing_Pipeline - Indicator normalisation and enrichment
4. Analysis_Team - Analyst capabilities, training, and workload
5. MITRE_Coverage - ATT&CK technique coverage and gap analysis
6. Quality_Assurance - Validation processes and accuracy metrics
7. Tool_Integration - SIEM/SOAR/security tool integrations
8. Performance_Metrics - SLA compliance, processing volumes, timeliness
9. Intelligence_Lifecycle - Aging, archival, and retention policies
10. Gap_Analysis - Workflow deficiencies and remediation plans
11. Evidence_Register - Audit evidence tracking
12. Metadata - Assessment metadata and approval workflow

**Key Features:**
- MITRE ATT&CK technique coverage heatmap with gap visualization
- CVSS enrichment and validation workflow tracking (VTL integration)
- Analyst workload analysis and capacity planning
- Indicator processing SLA compliance monitoring
- Data validation with industry-standard dropdown lists
- Conditional formatting for workflow health status
- Protected formulas with unprotected input cells
- Integration with A.5.7.1 (source utilization) and A.5.7.3 (distribution)
- Evidence linkage for audit traceability

**VulnerabilityThreatLink (VTL) Integration:**
This assessment tracks CVSS enrichment and validation workflows that ensure
threat intelligence includes accurate severity ratings. Quality CVSS data
from collection/analysis enables effective vulnerability-threat correlation
and risk-based patching decisions in Control A.8.8.

**Integration:**
This assessment feeds into:
- A.5.7.4 Threat Intelligence Effectiveness Dashboard (workflow metrics)
- A.5.7.1 Sources Assessment (source utilization and effectiveness)
- A.5.7.3 Integration & Distribution (consumer requirements)
- A.8.8 Vulnerability Management (CVSS quality for patch prioritization)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a57_2_collection_analysis.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a57_2_collection_analysis.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a57_2_collection_analysis.py --date 20250115

Output:
    File: ISMS_A_5_7_2_Collection_Analysis_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)
    Sheets: 12 worksheets (see structure above)

Post-Generation Steps:
    1. Document all automated collection workflows (API integrations, feeds)
    2. Map manual collection processes and analyst responsibilities
    3. Assess indicator processing pipeline and normalisation quality
    4. Evaluate analyst team capabilities, training, and workload
    5. Complete MITRE ATT&CK coverage analysis and gap identification
    6. Review quality assurance processes and validation metrics
    7. Verify tool integrations (SIEM, SOAR, ticketing, etc.)
    8. Measure performance against SLAs (timeliness, accuracy, volume)
    9. Define intelligence lifecycle policies (aging, archival, retention)
    10. Identify workflow gaps and create remediation plans
    11. Collect audit evidence (logs, metrics, validation reports)
    12. Obtain stakeholder approvals
    13. Feed results into A.5.7.4 Effectiveness Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.7
Assessment Domain:    2 of 5 (Intelligence Collection & Analysis Workflows)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.7: Threat Intelligence Policy (Governance)
    - ISMS-IMP-A.5.7.2: Intelligence Collection & Analysis Implementation Guide
    - ISMS-IMP-A.5.7.1: Threat Intelligence Sources Assessment (Domain 1)
    - ISMS-IMP-A.5.7.3: Intelligence Integration & Distribution Assessment (Domain 3)
    - ISO 27002:2022 Implementation Guidance for Control A.5.7
    - MITRE ATT&CK Framework (Enterprise v15.1)
    - STIX/TAXII Standards for Threat Intelligence Exchange
    - CVSS Scoring Standard v4.0 / v3.1 Specifications

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
- Expanded MITRE ATT&CK coverage analysis (Enterprise v15.1)
    - Added CVSS enrichment and validation workflow tracking (VTL integration)
    - Enhanced quality assurance metrics (confidence levels, false positives)
    - Added analyst workload and capacity planning analysis
    - Improved integration tracking with SIEM/SOAR platforms
    - Added intelligence lifecycle management assessment
    - Enhanced audit evidence collection mechanisms
    - Updated conditional formatting for workflow health visualization

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Intelligence Quality Standards:**
Collection and analysis workflows directly impact intelligence quality.
Establish clear quality metrics:
- Accuracy: Validation rate, false positive/negative percentages
- Timeliness: Time from source ingestion to analyst review to distribution
- Completeness: CVSS enrichment rate, context completeness, attribution quality
- Consistency: Normalisation success rate, schema compliance
- Actionability: Percentage of intelligence leading to security actions

Poor workflow quality degrades downstream security operations effectiveness.

**MITRE ATT&CK Coverage (AUDIT CRITICAL):**
Auditors will verify that threat intelligence collection covers relevant
techniques based on your threat model. Document:
- Coverage percentage by ATT&CK tactic (Reconnaissance through Impact)
- Justification for gaps (techniques irrelevant to your environment)
- Plans to address coverage deficiencies
- Regular review cycle (quarterly recommended)

Incomplete coverage may indicate blind spots in threat detection capabilities.

**Analyst Workload Management:**
Unsustainable analyst workload leads to:
- Decreased analysis quality (rushed assessments)
- Increased analyst burnout and turnover
- Processing backlogs and SLA violations
- Missed critical intelligence

Monitor analyst productivity metrics and adjust staffing/automation accordingly.
Target: <40 hours/week analysis workload per analyst for sustainable operations.

**CVSS Enrichment Workflow (VTL Critical):**
For threat intelligence lacking CVSS scores, document enrichment processes:
- Who performs CVSS scoring (analyst roles, training)
- Validation against authoritative sources (NVD, vendor advisories)
- Quality assurance processes (peer review, spot checks)
- Accuracy targets (>90% agreement with authoritative scores)

CVSS enrichment quality directly impacts vulnerability management automation.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will verify:
- Documented collection workflows (automated and manual)
- Analyst competency and training records
- Quality assurance processes and validation results
- Integration with security operations workflows
- Performance metrics and SLA compliance

Maintain evidence in Sheet 11 (Evidence_Register) with proper linkage.

**Data Protection:**
Assessment workbooks contain sensitive operational details:
- Threat intelligence workflows and analysis procedures
- Analyst identities and capabilities assessment
- Tool integration technical details
- Performance metrics revealing operational gaps

Classification: CONFIDENTIAL - Internal Use Only. Restrict to security team.

**Maintenance:**
Review and update assessment:
- Monthly: Performance metrics and SLA compliance
- Quarterly: MITRE ATT&CK coverage review and analyst workload analysis
- Semi-annually: Workflow optimization and tool integration review
- Annually: Complete reassessment of collection and analysis capabilities
- Ad-hoc: When workflows change or new tools are deployed

**Quality Assurance:**
Have senior threat intelligence analysts and SOC leadership validate
assessments before using results for resource allocation or workflow changes.
Cross-reference with A.5.7.1 (source effectiveness) and A.5.7.3 (consumer needs).

**Regulatory Alignment:**
Some regulations require specific threat intelligence capabilities:
- Financial sector: Fraud and financial crime intelligence workflows
- Healthcare: Healthcare-specific threat intelligence (medical device, EHR)
- Critical infrastructure: ICS/SCADA threat intelligence processes
- Government: Classified intelligence handling procedures (if applicable)

Customize workflow assessments to include regulatory-specific requirements.

**Tool Integration Dependencies:**
Collection and analysis effectiveness depends heavily on tool integrations:
- SIEM: Automated indicator ingestion and correlation
- SOAR: Playbook automation for intelligence-driven response
- Threat intelligence platform (TIP): Central intelligence repository
- Vulnerability scanners: CVSS correlation for patch prioritization
- EDR/XDR: Endpoint indicator deployment and hunting

Poor integrations create manual workflows that don't scale. Prioritize automation.

**MITRE ATT&CK Framework Usage:**
This assessment uses MITRE ATT&CK Enterprise Framework v15.1 (or current).
Update technique coverage mapping when new ATT&CK versions are released.
Focus coverage analysis on tactics/techniques relevant to your threat model.

**Business Impact:**
Inadequate collection and analysis workflows result in:
- Delayed threat detection and incident response
- Missed threat actor campaigns targeting your sector
- Inability to prioritize vulnerabilities based on exploitation likelihood
- Reduced security ROI from threat intelligence investments

This assessment helps quantify and remediate workflow deficiencies.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.7.2"
WORKBOOK_NAME = "Intelligence Collection & Analysis Assessment"
CONTROL_ID = "A.5.7"
CONTROL_NAME = "Threat Intelligence"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure: 15 sheets (Gold Standard tail order: ...data → ER → SD → AS)
    sheets = [
        "Instructions & Legend",
        "Intelligence Requirements",     # PIR/SIR/IR tracking
        "Collection Sources",            # Source documentation
        "Raw Intelligence Log",          # Raw intel tracking
        "Intelligence Production",
        "MITRE Mapping",
        "Quality Metrics",
        "Process Maturity",              # Capability maturity assessment
        "Action Items",
        "Analysis Tools",
        "Threat Actor Profiles",
        "Campaign Tracking",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles



_STYLES = setup_styles()
def get_border():
    """Create new border object (avoid shared object warnings)."""
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATION FUNCTIONS
# ============================================================================

def setup_validations():
    """Create all data validation dropdowns."""
    validations = {}

    # Intelligence Type
    validations['intelligence_type'] = DataValidation(
        type="list",
        formula1='"Strategic,Tactical,Operational,Technical"',
        allow_blank=False
    )
    
    # Priority
    validations['priority'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    
    # Collection Method
    validations['collection_method'] = DataValidation(
        type="list",
        formula1='"Automated_Feed,Manual_Research,OSINT,Partnership,Internal_Telemetry"',
        allow_blank=False
    )
    
    # Collection Frequency
    validations['collection_frequency'] = DataValidation(
        type="list",
        formula1='"Real-time,Hourly,Daily,Weekly,On_Demand"',
        allow_blank=False
    )
    
    # Framework Type
    validations['framework_type'] = DataValidation(
        type="list",
        formula1='"Threat_Model,Analysis_Method,Taxonomy,Tool"',
        allow_blank=False
    )
    
    # Use Case
    validations['use_case'] = DataValidation(
        type="list",
        formula1='"Threat_Profiling,IOC_Analysis,Campaign_Tracking,Reporting,Integration"',
        allow_blank=False
    )
    
    # Implementation Status
    validations['implementation_status'] = DataValidation(
        type="list",
        formula1='"Fully_Implemented,Partially_Implemented,Planned,Not_Used"',
        allow_blank=False
    )
    
    # Yes/No/Optional
    validations['yes_no_optional'] = DataValidation(
        type="list",
        formula1='"Yes,No,Optional"',
        allow_blank=False
    )
    
    # Effectiveness Rating
    validations['effectiveness_rating'] = DataValidation(
        type="list",
        formula1='"Excellent,Good,Fair,Poor"',
        allow_blank=True
    )
    
    # Role
    validations['analyst_role'] = DataValidation(
        type="list",
        formula1='"Lead_Analyst,Senior_Analyst,Analyst,Junior_Analyst,Intern"',
        allow_blank=False
    )
    
    # Primary Focus Area
    validations['focus_area'] = DataValidation(
        type="list",
        formula1='"Strategic,Tactical,Operational,Technical"',
        allow_blank=False
    )
    
    # Skill Levels
    validations['skill_level'] = DataValidation(
        type="list",
        formula1='"Expert,Advanced,Intermediate,Beginner,None"',
        allow_blank=False
    )
    
    # Product Type
    validations['product_type'] = DataValidation(
        type="list",
        formula1='"Strategic_Report,Tactical_Brief,Operational_Alert,Technical_Analysis,IOC_Package,Threat_Profile"',
        allow_blank=False
    )
    
    # Product Status
    validations['product_status'] = DataValidation(
        type="list",
        formula1='"Draft,Review,Published,Archived"',
        allow_blank=False
    )
    
    # Confidence Level
    validations['confidence_level'] = DataValidation(
        type="list",
        formula1='"High,Medium,Low,Unconfirmed"',
        allow_blank=False
    )
    
    # TLP Classification
    validations['tlp'] = DataValidation(
        type="list",
        formula1='"TLP:CLEAR,TLP:GREEN,TLP:AMBER,TLP:AMBER+STRICT,TLP:RED"',
        allow_blank=False
    )
    
    # MITRE Coverage
    validations['mitre_coverage'] = DataValidation(
        type="list",
        formula1='"Full,Partial,Limited,None"',
        allow_blank=False
    )
    
    # Threat Actor Type (VTL schema)
    validations['threat_actor_type'] = DataValidation(
        type="list",
        formula1='"Nation-State,Criminal_Syndicate,Hacktivist,Insider_Threat,Script_Kiddie,Unknown"',
        allow_blank=True
    )
    
    # Exploitation Status (VTL schema - CRITICAL)
    validations['exploitation_status'] = DataValidation(
        type="list",
        formula1='"Not_Exploited,PoC_Exists,Active_Exploitation,Widespread"',
        allow_blank=False
    )
    
    # Remediation Urgency (VTL schema)
    validations['remediation_urgency'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,Info"',
        allow_blank=False
    )
    
    # Remediation Status (VTL schema)
    validations['remediation_status'] = DataValidation(
        type="list",
        formula1='"Open,In_Progress,Patched,Mitigated,Risk_Accepted,Verified"',
        allow_blank=False
    )
    
    # Yes/No
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    
    # Maturity Levels (1-5 CMM)
    validations['maturity_level'] = DataValidation(
        type="list",
        formula1='"1 - Initial,2 - Managed,3 - Defined,4 - Quantitatively Managed,5 - Optimizing"',
        allow_blank=False
    )
    
    # Action Status
    validations['action_status'] = DataValidation(
        type="list",
        formula1='"Open,In_Progress,Blocked,Resolved,Closed"',
        allow_blank=False
    )
    

    # NEW v1.0 validations for Sheets 11-13
    validations['tool_category'] = DataValidation(
        type="list",
        formula1='"TIP,SIEM,Malware_Analysis,OSINT,Collaboration,Visualization,Scripting,Other"',
        allow_blank=True
    )
    validations['license_type'] = DataValidation(
        type="list",
        formula1='"Commercial,Open_Source,Freeware,Internal_Developed"',
        allow_blank=True
    )
    validations['yes_no_na'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    validations['yes_no_unknown'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    validations['targeting_status'] = DataValidation(
        type="list",
        formula1='"Confirmed,Suspected,No,Unknown"',
        allow_blank=True
    )
    validations['motivation'] = DataValidation(
        type="list",
        formula1='"Espionage,Financial,Disruption,Ideology,Unknown"',
        allow_blank=True
    )
    validations['sophistication'] = DataValidation(
        type="list",
        formula1='"Advanced,Moderate,Low"',
        allow_blank=True
    )
    validations['campaign_status'] = DataValidation(
        type="list",
        formula1='"Active,Concluded,Dormant,Unknown"',
        allow_blank=True
    )
    validations['campaign_objective'] = DataValidation(
        type="list",
        formula1='"Espionage,Data_Theft,Disruption,Ransomware,Credential_Harvesting,Other"',
        allow_blank=True
    )
    validations['threat_level'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    validations['monitoring_status'] = DataValidation(
        type="list",
        formula1='"Active_Monitoring,Passive_Monitoring,Concluded"',
        allow_blank=True
    )

    validations['cvss_version'] = DataValidation(
        type="list",
        formula1='"4.0,3.1"',
        allow_blank=True
    )
    

    # Validations for Sheet 11 (Analysis_Tools)
    validations['tool_category'] = DataValidation(
        type="list",
        formula1='"TIP,SIEM,Malware_Analysis,OSINT,Collaboration,Visualization,Scripting,Other"',
        allow_blank=True
    )
    validations['license_type'] = DataValidation(
        type="list",
        formula1='"Commercial,Open_Source,Freeware,Internal_Developed"',
        allow_blank=True
    )
    validations['integration_status'] = DataValidation(
        type="list",
        formula1='"Integrated,Standalone,Planned,Deprecated"',
        allow_blank=True
    )
    validations['yes_no_na'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    
    # Validations for Sheet 12 (Threat_Actor_Profiles)
    validations['yes_no_unknown'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    validations['targeting_status'] = DataValidation(
        type="list",
        formula1='"Confirmed,Suspected,No,Unknown"',
        allow_blank=True
    )
    validations['motivation'] = DataValidation(
        type="list",
        formula1='"Espionage,Financial,Disruption,Ideology,Unknown"',
        allow_blank=True
    )
    validations['sophistication'] = DataValidation(
        type="list",
        formula1='"Advanced,Moderate,Low"',
        allow_blank=True
    )
    
    # Validations for Sheet 13 (Campaign_Tracking)
    validations['campaign_status'] = DataValidation(
        type="list",
        formula1='"Active,Concluded,Dormant,Unknown"',
        allow_blank=True
    )
    validations['campaign_objective'] = DataValidation(
        type="list",
        formula1='"Espionage,Data_Theft,Disruption,Ransomware,Credential_Harvesting,Other"',
        allow_blank=True
    )
    validations['threat_level'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    validations['monitoring_status'] = DataValidation(
        type="list",
        formula1='"Active_Monitoring,Passive_Monitoring,Concluded"',
        allow_blank=True
    )
    

    return validations


# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS
# ============================================================================

def create_instructions(ws, styles):
    """Create GS-IL-compliant Instructions & Legend sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 2 — empty

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Workbook Title", WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Instructions section
    row = 12
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Define intelligence requirements (PIRs/SIRs) in the Intelligence Requirements sheet — align each requirement to a specific threat scenario, business process, or risk.",
        "2. Map collection activities to each requirement in the Collection Plan sheet — identify responsible owners, collection methods, sources, and frequency.",
        "3. Record processed intelligence findings in the Analysis Records sheet — document source, confidence level, threat actor attribution, and recommended defensive actions.",
        "4. Document all intelligence sharing activities in the Sharing Records sheet — include recipient organisation, format, TLP classification, and date shared.",
        "5. Apply TLP labels (TLP:RED, TLP:AMBER, TLP:GREEN, TLP:CLEAR) to all intelligence products before dissemination to internal or external recipients.",
        "6. Ensure all shared intelligence includes provenance information, confidence ratings, and handling restrictions as required by your sharing agreements.",
        "7. Maintain the Evidence Register with analysis reports, intelligence products, sharing agreements, and dissemination logs.",
        "8. Obtain sign-off from the Threat Intelligence Manager, Information Security Officer, and Legal (for external sharing arrangements).",
    ]
    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # Status Legend section
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("✓", "Compliant / Complete", "Requirement fully met", _green),
        ("⚠", "Partial / In Progress", "Partially met or in progress", _amber),
        ("✗", "Non-Compliant / Not Started", "Requirement not met", _red),
        ("—", "Not Applicable", "Not applicable to this assessment", None),
    ]
    row += 1
    for sym, status, desc, fill in legend_rows:
        ws.cell(row=row, column=1, value=sym).border = _border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    # Acceptable Evidence section
    row += 1
    ws[f"A{row}"] = "Acceptable Evidence (examples)"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for evidence in [
        "\u2713 Intelligence requirements documentation (PIR / SIR register)",
        "\u2713 Intelligence collection and analysis reports",
        "\u2713 TLP-marked intelligence products disseminated internally or externally",
        "\u2713 Sharing agreements with ISACs, sector partners, or government bodies",
        "\u2713 Dissemination records and recipient acknowledgement logs",
    ]:
        ws[f"A{row}"] = evidence
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"
def create_intelligence_requirements(ws, styles, validations):
    """Sheet 2: Intelligence_Requirements - PIR/SIR/IR tracking per IMP spec."""

    # Title
    ws.merge_cells("A1:R1")
    ws["A1"] = "INTELLIGENCE REQUIREMENTS (PIR/SIR/IR)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:R2")
    ws["A2"] = "Document all intelligence requirements (PIRs, SIRs, IRs). Map sources from A.5.7.1. Identify coverage gaps."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers (Row 4)
    headers = [
        "Requirement ID", "Intelligence Type", "Requirement Description", "Priority",
        "Target Sector", "Target Region", "Threat Category", "Collection Source 1",
        "Collection Source 2", "Collection Source 3", "Collection Method",
        "Coverage Status", "Collection Frequency", "Last Collected", "Gap Identified",
        "Gap Remediation", "Responsible Analyst", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_req = [
        "REQ-2025-EXX", "Strategic",
        "Track APT groups targeting EMEA financial sector", "Critical",
        "Finance", "EMEA", "APT/Ransomware",
        "MISP Community", "Commercial TI", "OSINT", "Automated API",
        "Adequate", "Weekly", "25.01.2026", "No", "", "J. Smith", "",
    ]
    for col_idx, val in enumerate(sample_req, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Requirement_ID (auto-formula — empty when no data)
        cell = ws.cell(row=row, column=1)
        cell.value = f'=IF(B{row}<>"","REQ-"&TEXT(ROW()-5,"000"),"")'
        cell.fill = _grey_d
        cell.border = get_border()

        # Intelligence_Type
        cell = ws.cell(row=row, column=2)
        cell.fill = _yell
        cell.border = get_border()
        validations['intelligence_type'].add(cell)

        # Requirement_Description
        cell = ws.cell(row=row, column=3)
        cell.fill = _yell
        cell.border = get_border()

        # Priority
        cell = ws.cell(row=row, column=4)
        cell.fill = _yell
        cell.border = get_border()
        validations['priority'].add(cell)

        # Target_Sector, Target_Region, Threat_Category (text input)
        for col in [5, 6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()

        # Collection_Source_1, 2, 3 (text input - reference to 5.7.1)
        for col in [8, 9, 10]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()

        # Collection_Method
        cell = ws.cell(row=row, column=11)
        cell.fill = _yell
        cell.border = get_border()
        validations['collection_method'].add(cell)

        # Coverage_Status (formula)
        cell = ws.cell(row=row, column=12)
        cell.value = f'=IF(COUNTA(H{row}:J{row})>=2,"Adequate",IF(COUNTA(H{row}:J{row})=1,"Minimal","Gap"))'
        cell.fill = _grey_d
        cell.border = get_border()

        # Collection_Frequency
        cell = ws.cell(row=row, column=13)
        cell.fill = _yell
        cell.border = get_border()
        validations['collection_frequency'].add(cell)

        # Last_Collected (date)
        cell = ws.cell(row=row, column=14)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Gap_Identified (formula)
        cell = ws.cell(row=row, column=15)
        cell.value = f'=IF(L{row}="Gap","Yes","No")'
        cell.fill = _grey_d
        cell.border = get_border()

        # Gap_Remediation, Responsible_Analyst, Notes
        for col in [16, 17, 18]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()

    # Apply validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Conditional formatting (skip sample row: L6:L55)
    ws.conditional_formatting.add(
        'L6:L55',
        CellIsRule(operator='equal', formula=['"Gap"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'L6:L55',
        CellIsRule(operator='equal', formula=['"Minimal"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'L6:L55',
        CellIsRule(operator='equal', formula=['"Adequate"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    # Summary section (row 57 — after sample row 5 + 50 data rows 6-55)
    ws.merge_cells("A57:D57")
    ws["A57"] = "COLLECTION COVERAGE SUMMARY"
    ws["A57"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A57"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    summary_metrics = [
        ("Total Requirements:", '=COUNTA(C6:C55)'),
        ("Critical Priority:", '=COUNTIF(D6:D55,"Critical")'),
        ("High Priority:", '=COUNTIF(D6:D55,"High")'),
        ("Coverage Gaps:", '=COUNTIF(L6:L55,"Gap")'),
        ("Minimal Coverage:", '=COUNTIF(L6:L55,"Minimal")'),
        ("Adequate Coverage:", '=COUNTIF(L6:L55,"Adequate")'),
    ]

    row = 58
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    for col in ['H', 'I', 'J']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 14
    ws.column_dimensions['P'].width = 30
    ws.column_dimensions['Q'].width = 25
    ws.column_dimensions['R'].width = 30
    
    ws.freeze_panes = "A4"


def create_collection_sources(ws, styles, validations):
    """Sheet 3: Collection_Sources - Document intelligence collection sources per IMP spec."""

    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "COLLECTION SOURCES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Document collection workflows for each source from A.5.7.1. Map data formats, update frequency, and integration status."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column headers per IMP specification
    headers = [
        "Source ID", "Source Name", "Source Type", "Data Format", "Update Frequency",
        "Coverage Geographic", "Coverage Sector", "Coverage Threat Types",
        "Integration Status", "Integration Platform", "Cost Annual", "Contract Expiry",
        "Primary Contact", "Data Quality Rating", "Last Review Date", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_src = [
        "SRC-001", "MISP Community", "OSINT", "STIX/TAXII", "Daily",
        "Global", "Finance/Tech", "APT/Ransomware", "Fully_Implemented",
        "REST API", "", "", "J. Smith", "Good", "25.01.2026", "",
    ]
    for col_idx, val in enumerate(sample_src, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Source_ID (col 1)
        cell = ws.cell(row=row, column=1)
        cell.fill = _grey_d
        cell.border = get_border()

        # Source_Name (col 2)
        cell = ws.cell(row=row, column=2)
        cell.fill = _yell
        cell.border = get_border()

        # Source_Type (col 3)
        cell = ws.cell(row=row, column=3)
        cell.fill = _yell
        cell.border = get_border()

        # Data_Format (col 4)
        cell = ws.cell(row=row, column=4)
        cell.fill = _yell
        cell.border = get_border()

        # Update_Frequency (col 5)
        cell = ws.cell(row=row, column=5)
        cell.fill = _yell
        cell.border = get_border()
        validations['collection_frequency'].add(cell)

        # Coverage_Geographic (col 6)
        cell = ws.cell(row=row, column=6)
        cell.fill = _yell
        cell.border = get_border()

        # Coverage_Sector (col 7)
        cell = ws.cell(row=row, column=7)
        cell.fill = _yell
        cell.border = get_border()

        # Coverage_Threat_Types (col 8)
        cell = ws.cell(row=row, column=8)
        cell.fill = _yell
        cell.border = get_border()

        # Integration_Status (col 9)
        cell = ws.cell(row=row, column=9)
        cell.fill = _yell
        cell.border = get_border()
        validations['implementation_status'].add(cell)

        # Integration_Platform (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = _yell
        cell.border = get_border()

        # Cost_Annual (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = '#,##0.00'

        # Contract_Expiry (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Primary_Contact (col 13)
        cell = ws.cell(row=row, column=13)
        cell.fill = _yell
        cell.border = get_border()

        # Data_Quality_Rating (col 14)
        cell = ws.cell(row=row, column=14)
        cell.fill = _yell
        cell.border = get_border()
        validations['effectiveness_rating'].add(cell)

        # Last_Review_Date (col 15)
        cell = ws.cell(row=row, column=15)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Notes (col 16)
        cell = ws.cell(row=row, column=16)
        cell.fill = _yell
        cell.border = get_border()

    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Summary section (row 57 — after sample + 50 data rows)
    ws.merge_cells("A57:D57")
    ws["A57"] = "COLLECTION SOURCES SUMMARY"
    ws["A57"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A57"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    summary = [
        ("Total Sources:", '=COUNTA(B6:B55)'),
        ("Integrated Sources:", '=COUNTIF(I6:I55,"Fully_Implemented")'),
        ("Manual Sources:", '=COUNTIF(I6:I55,"Partially_Implemented")'),
        ("Excellent Quality:", '=COUNTIF(N6:N55,"Excellent")'),
        ("Good Quality:", '=COUNTIF(N6:N55,"Good")'),
    ]

    row = 58
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 14
    ws.column_dimensions['P'].width = 35

    ws.freeze_panes = "A4"


def create_analyst_capabilities(ws, styles, validations):
    """Sheet 4: Analyst Capabilities - Skills matrix and training."""
    
    # Title
    ws.merge_cells("A1:U1")
    ws["A1"] = "ANALYST CAPABILITIES & SKILLS MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:U2")
    ws["A2"] = "Skills assessment for all TI analysts. Track expertise levels, certifications, and training plans."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers
    headers = [
        "Analyst ID", "Analyst Name", "Email", "Role", "Team", "Years Experience TI",
        "Years Experience InfoSec", "Primary Focus Area", "Secondary Focus Area",
        "Skill OSINT", "Skill Malware Analysis", "Skill MITRE ATTACK", "Skill Scripting Python",
        "Skill Threat Modeling", "Skill Report Writing", "Skill Briefing Presentation",
        "Certifications", "Training Planned 2025", "Training Budget 2025",
        "Last Performance Review", "Capability Gap Identified", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_ana = [
        "ANALYST-001", "J. Smith", "j.smith@org.com", "Senior Analyst",
        "Threat Intel", "5", "8", "APT Analysis", "MITRE ATT&CK",
        "Expert", "Advanced", "Expert", "Intermediate", "Advanced", "Expert", "Advanced",
        "GIAC GREM", "MITRE ATT&CK Practitioner", "5000", "25.01.2026", "No", "",
    ]
    for col_idx, val in enumerate(sample_ana, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Analyst_ID
        cell = ws.cell(row=row, column=1)
        cell.fill = _grey_d
        cell.border = get_border()

        # Analyst_Name, Email, Team
        for col in [2, 3, 5]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()

        # Role
        cell = ws.cell(row=row, column=4)
        cell.fill = _yell
        cell.border = get_border()
        validations['analyst_role'].add(cell)

        # Years experience (numbers)
        for col in [6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()
            cell.number_format = '0'

        # Primary/Secondary Focus Area
        for col in [8, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()
            validations['focus_area'].add(cell)

        # Skill levels (columns 10-16)
        for col in range(10, 17):
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()
            validations['skill_level'].add(cell)

        # Certifications, Training Planned, Capability Gap, Notes
        for col in [17, 18, 21, 22]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()

        # Training Budget
        cell = ws.cell(row=row, column=19)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = '#,##0.00'

        # Last Performance Review
        cell = ws.cell(row=row, column=20)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Skill Level Reference Table (row 57 — after sample + 50 data rows)
    ws.merge_cells("A57:F57")
    ws["A57"] = "SKILL LEVEL REFERENCE"
    ws["A57"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A57"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    skill_ref = [
        ("Expert", "Can teach others, industry-recognised expertise"),
        ("Advanced", "Fully autonomous, handles complex scenarios"),
        ("Intermediate", "Autonomous for routine tasks, guidance for complex"),
        ("Beginner", "Requires supervision and guidance"),
        ("None", "No capability in this area"),
    ]

    row = 58
    for level, desc in skill_ref:
        ws[f"A{row}"] = level
        ws[f"B{row}"] = desc
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].font = Font(name="Calibri", size=10)
        row += 1

    # Team Summary (row 64)
    ws.merge_cells("A64:D64")
    ws["A64"] = "TEAM CAPABILITY SUMMARY"
    ws["A64"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A64"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary = [
        ("Total Analysts:", '=COUNTA(B6:B55)'),
        ("Avg Years TI Experience:", '=AVERAGE(F6:F55)'),
        ("Total Training Budget 2025:", '=SUM(S6:S55)'),
        ("Certifications Held:", '=COUNTA(Q6:Q55)'),
    ]

    row = 65
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        if "Budget" in label:
            ws[f"B{row}"].number_format = '#,##0.00'
        elif "Avg Years" in label:
            ws[f"B{row}"].number_format = '0.0'
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    for col in ['F', 'G']:
        ws.column_dimensions[col].width = 18
    for col in ['H', 'I']:
        ws.column_dimensions[col].width = 18
    for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['Q'].width = 30
    ws.column_dimensions['R'].width = 30
    ws.column_dimensions['S'].width = 18
    ws.column_dimensions['T'].width = 18
    ws.column_dimensions['U'].width = 30
    ws.column_dimensions['V'].width = 35
    
    ws.freeze_panes = "A4"


def create_intelligence_production(ws, styles, validations):
    """Sheet 5: Intelligence Production - Track intelligence products."""
    
    # Title
    ws.merge_cells("A1:O1")
    ws["A1"] = "INTELLIGENCE PRODUCTION TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:O2")
    ws["A2"] = "Track all intelligence products (reports, briefings, alerts, IOC packages). Measure quality and stakeholder feedback."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers
    headers = [
        "Product ID", "Product Type", "Product Title", "Primary Author", "Contributors",
        "Publication Date", "TLP Classification", "Confidence Level", "Target Audience",
        "Stakeholder Feedback Rating", "Consumption Rate", "Actionability Score",
        "Time to Produce Hours", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_prod = [
        "PROD-2026-EXX", "Tactical Report", "APT29 EMEA Campaign Briefing",
        "J. Smith", "TI Team", "25.01.2026", "TLP:AMBER", "High",
        "SOC / IR Team", "", "", "", "", "Published", "",
    ]
    for col_idx, val in enumerate(sample_prod, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Product_ID
        cell = ws.cell(row=row, column=1)
        cell.fill = _grey_d
        cell.border = get_border()
        
        # Product_Type
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['product_type'].add(cell)
        
        # Product_Title, Primary_Author, Contributors, Target_Audience
        for col in [3, 4, 5, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Publication_Date
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # TLP_Classification
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['tlp'].add(cell)
        
        # Confidence_Level
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['confidence_level'].add(cell)
        
        # Stakeholder_Feedback_Rating (1-5)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0'
        
        # Consumption_Rate (percentage)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0%'
        
        # Actionability_Score (1-5)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0'
        
        # Time_to_Produce_Hours
        cell = ws.cell(row=row, column=13)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0.0'
        
        # Status
        cell = ws.cell(row=row, column=14)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['product_status'].add(cell)
        
        # Notes
        cell = ws.cell(row=row, column=15)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
    
    # Apply only the validations used on this sheet (avoid shared-ref contamination)
    for key in ['product_type', 'tlp', 'confidence_level', 'product_status']:
        ws.add_data_validation(validations[key])

    # Summary section (row 57 — after sample + 50 data rows)
    ws.merge_cells("A57:D57")
    ws["A57"] = "PRODUCTION SUMMARY"
    ws["A57"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A57"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    summary = [
        ("Total Products:", '=COUNTA(C6:C55)'),
        ("Published This Quarter:", '=COUNTIF(N6:N55,"Published")'),
        ("Avg Feedback Rating:", '=AVERAGE(J6:J55)'),
        ("Avg Consumption Rate:", '=AVERAGE(K6:K55)'),
        ("Avg Time to Produce (hrs):", '=AVERAGE(M6:M55)'),
    ]

    row = 58
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        if "Rate" in label:
            ws[f"B{row}"].number_format = '0%'
        elif "Rating" in label or "hrs" in label:
            ws[f"B{row}"].number_format = '0.0'
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 30
    
    ws.freeze_panes = "A4"
# ============================================================================
# SECTION 3: SHEETS 6-8 (MITRE MAPPING, QUALITY METRICS, VTL SCHEMA)
# ============================================================================

def create_mitre_mapping(ws, styles, validations):
    """Sheet 6: MITRE ATT&CK Mapping - Technique coverage analysis."""
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "MITRE ATT&CK TECHNIQUE COVERAGE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Assess threat intelligence coverage of MITRE ATT&CK techniques. Focus on techniques relevant to your threat landscape."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers
    headers = [
        "Technique ID", "Tactic", "Technique Name", "Sub Technique", "Priority for Org",
        "Coverage Level", "Intelligence Sources", "Last Intelligence Update",
        "Detection Capability", "Use Cases Identified", "Gap Analysis", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # MAX-004 FIX: Reduced from 100 FFFFCC rows (5-104) to 1 F2F2F2 sample row (row 5)
    # + 50 FFFFCC empty rows (rows 6-55).
    # Sample row uses T1190 as a representative format guide example.
    _grey_fill = PatternFill("solid", fgColor="F2F2F2")
    _ffffcc_fill = PatternFill("solid", fgColor="FFFFCC")

    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    sample_data = ["T1190", "Initial Access", "Exploit Public-Facing Application",
                   "", "Critical", "Partial", "OSINT Feed, Commercial TI",
                   "", "Yes", "Phishing, Malware delivery", "Coverage gap on sub-techniques", ""]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey_fill
        cell.border = get_border()
        if val:
            cell.value = val
        if col_idx == 8:
            cell.number_format = 'DD.MM.YYYY'

    # Rows 6-55: 50 FFFFCC empty data rows
    for row in range(6, 56):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = _ffffcc_fill
            cell.border = get_border()

            if col == 5:  # Priority
                validations['priority'].add(cell)
            elif col == 6:  # Coverage_Level
                validations['mitre_coverage'].add(cell)
            elif col == 8:  # Last_Intelligence_Update
                cell.number_format = 'DD.MM.YYYY'
            elif col == 9:  # Detection_Capability
                validations['yes_no'].add(cell)

    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Conditional formatting for coverage levels (F6:F55 — excludes sample row)
    ws.conditional_formatting.add(
        'F6:F55',
        CellIsRule(operator='equal', formula=['"Full"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'F6:F55',
        CellIsRule(operator='equal', formula=['"None"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'F6:F55',
        CellIsRule(operator='equal', formula=['"Limited"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )

    # Summary section (shifted to row 57 after MAX-004 fix: was row 106)
    ws.merge_cells("A57:D57")
    ws["A57"] = "MITRE COVERAGE SUMMARY"
    ws["A57"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A57"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Updated COUNTA/COUNTIF ranges: A6:A55 / E6:E55 / F6:F55 (exclude sample row 5)
    summary = [
        ("Total Techniques Tracked:", '=COUNTA(A6:A55)'),
        ("Critical Priority:", '=COUNTIF(E6:E55,"Critical")'),
        ("Full Coverage:", '=COUNTIF(F6:F55,"Full")'),
        ("Partial Coverage:", '=COUNTIF(F6:F55,"Partial")'),
        ("Limited Coverage:", '=COUNTIF(F6:F55,"Limited")'),
        ("No Coverage:", '=COUNTIF(F6:F55,"None")'),
        ("With Detection Capability:", '=COUNTIF(I6:I55,"Yes")'),
    ]

    row = 58
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 30

    ws.freeze_panes = "A4"


def create_quality_metrics(ws, styles, validations):
    """Sheet 7: Quality Metrics - KPIs and performance tracking."""
    
    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "INTELLIGENCE QUALITY METRICS & KPIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "Track key performance indicators for intelligence collection and analysis operations."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers
    headers = [
        "KPI ID", "KPI Name", "Description", "Target Value", "Current Value",
        "Last Month Value", "Measurement Period", "Performance vs Target",
        "Trend", "Data Source", "Owner", "Last Updated", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Pre-populate standard KPIs
    kpis = [
        ("KPI-001", "Intelligence Products per Month", "Count of formal intelligence products", "20", "Sheet 5"),
        ("KPI-002", "Avg Time to Produce Intelligence (hrs)", "Hours from start to publication", "48", "Sheet 5"),
        ("KPI-003", "Stakeholder Feedback Rating (1-5)", "Average rating from consumers", "4.0", "Sheet 5"),
        ("KPI-004", "Intelligence Actionability Rate (%)", "% of products acted upon", "75%", "Sheet 5"),
        ("KPI-005", "Mean Time from Intelligence to Action (hrs)", "Hours from publication to action", "24", "Sheet 5"),
        ("KPI-006", "MITRE ATT&CK Coverage (%)", "% of critical techniques covered", "80%", "Sheet 6"),
        ("KPI-007", "False Positive Rate (%)", "% of incorrect/unusable intelligence", "5%", "Sheet 5"),
        ("KPI-008", "Analyst Training Completion Rate (%)", "% of planned training completed", "90%", "Sheet 4"),
        ("KPI-009", "Requirement Fulfillment Rate (%)", "% of PIRs with adequate coverage", "85%", "Sheet 2"),
        ("KPI-010", "VT Links Created per Month", "Vulnerability-threat linkages", "10", "Sheet 8"),
    ]
    
    row = 5
    for kpi_id, name, desc, target, source in kpis:
        ws[f"A{row}"] = kpi_id
        ws[f"B{row}"] = name
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = target
        ws[f"J{row}"] = source
        
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            if col in [1, 8, 9]:  # ID, Performance_vs_Target, Trend (formulas)
                cell.fill = styles["formula_cell"]["fill"]
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Performance_vs_Target formula
        ws[f"H{row}"] = f'=IF(E{row}="","",IF(D{row}="","N/A",(E{row}/D{row}-1)*100))'
        
        # Trend formula (simplified - compare current to last month)
        ws[f"I{row}"] = f'=IF(OR(E{row}="",F{row}=""),"N/A",IF(E{row}>F{row},"Improving",IF(E{row}<F{row},"Declining","Stable")))'
        
        row += 1
    
    # Additional rows for custom KPIs (rows 15-24)
    for row in range(15, 25):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            if col in [1, 8, 9]:  # Formulas
                cell.fill = styles["formula_cell"]["fill"]
                if col == 8:
                    cell.value = f'=IF(E{row}="","",IF(D{row}="","N/A",(E{row}/D{row}-1)*100))'
                elif col == 9:
                    cell.value = f'=IF(OR(E{row}="",F{row}=""),"N/A",IF(E{row}>F{row},"Improving",IF(E{row}<F{row},"Declining","Stable")))'
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            
            if col == 12:  # Last_Updated
                cell.number_format = 'DD.MM.YYYY'
    
    # Conditional formatting for performance
    ws.conditional_formatting.add(
        'H5:H24',
        CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'H5:H24',
        CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # Summary dashboard (row 26)
    ws.merge_cells("A26:D26")
    ws["A26"] = "KPI SUMMARY"
    ws["A26"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A26"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total KPIs Tracked:", '=COUNTA(A5:A24)'),
        ("KPIs On Target (>=100%):", '=COUNTIF(H5:H24,">=0")'),
        ("KPIs Below Target:", '=COUNTIF(H5:H24,"<0")'),
        ("Improving Trends:", '=COUNTIF(I5:I24,"Improving")'),
        ("Declining Trends:", '=COUNTIF(I5:I24,"Declining")'),
    ]
    
    row = 27
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 30
    
    ws.freeze_panes = "A4"


def create_raw_intelligence_log(ws, styles, validations):
    """Sheet 4: Raw Intelligence Log - Track raw intel before analysis."""

    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "RAW INTELLIGENCE LOG"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Track all raw intelligence received before analysis. Log source, receipt time, initial triage, and processing status."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column headers (Row 4)
    headers = [
        "Log ID", "Date Received", "Time Received", "Source ID", "Source Type",
        "Intelligence Type", "TLP Classification", "Initial Triage Priority",
        "Brief Description", "Raw Data Location", "Assigned Analyst",
        "Processing Status", "Date Processed", "Linked Product ID",
        "Quality Score", "Notes"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()

    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_raw = [
        "RAW-2026-EXX", "25.02.2026", "09:00", "SRC-001", "OSINT",
        "Tactical", "TLP:GREEN", "High",
        "Phishing campaign targeting SWIFT payment portals", "/evidence/raw/raw-2026-exx.json",
        "J. Smith", "Received", "", "", "3", "",
    ]
    for col_idx, val in enumerate(sample_raw, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Log_ID
        cell = ws.cell(row=row, column=1)
        cell.fill = _grey_d
        cell.border = get_border()

        # Date_Received
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Time_Received
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'HH:MM'

        # Source_ID, Source_Type (text input)
        for col in [4, 5]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()

        # Intelligence_Type
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['intelligence_type'].add(cell)

        # TLP_Classification
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['tlp'].add(cell)

        # Initial_Triage_Priority
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['priority'].add(cell)

        # Brief_Description, Raw_Data_Location, Assigned_Analyst
        for col in [9, 10, 11]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()

        # Processing_Status
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['product_status'].add(cell)

        # Date_Processed
        cell = ws.cell(row=row, column=13)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Linked_Product_ID
        cell = ws.cell(row=row, column=14)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Quality_Score (1-5)
        cell = ws.cell(row=row, column=15)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Notes
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

    # Apply validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Conditional formatting for priority
    ws.conditional_formatting.add(
        'H6:H55',
        CellIsRule(operator='equal', formula=['"Critical"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'H6:H55',
        CellIsRule(operator='equal', formula=['"High"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )

    # Summary metrics (row 57 — after sample + 50 data rows)
    ws.merge_cells("A57:D57")
    ws["A57"] = "RAW INTELLIGENCE LOG SUMMARY"
    ws["A57"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A57"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    summary_metrics = [
        ("Total Raw Intel Logged:", '=COUNTA(I6:I55)'),
        ("Critical Priority:", '=COUNTIF(H6:H55,"Critical")'),
        ("High Priority:", '=COUNTIF(H6:H55,"High")'),
        ("Pending Processing:", '=COUNTIF(L6:L55,"Received")'),
        ("Processed:", '=COUNTIF(L6:L55,"Published")'),
        ("TLP:RED Items:", '=COUNTIF(G6:G55,"TLP:RED")'),
    ]

    row = 58
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 35

    ws.freeze_panes = "A4"


def create_process_maturity(ws, styles, validations):
    """Sheet 9: Process Maturity - CMM assessment."""
    
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "PROCESS MATURITY ASSESSMENT (CMM MODEL)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:G2")
    ws["A2"] = "Assess capability maturity using 5-level CMM model: 1-Initial, 2-Managed, 3-Defined, 4-Quantitatively Managed, 5-Optimizing"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers
    headers = [
        "Process Area", "Current Level", "Target Level", "Gap", "Evidence", "Action Plan", "Target Date"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Pre-populate standard process areas
    process_areas = [
        "Collection Planning",
        "Analysis Methodologies",
        "Production Quality",
        "Dissemination",
        "Feedback Mechanisms",
        "Integration (Technical)",
        "Integration (Control 8.8)",
        "Staff Training",
        "Metrics & KPIs",
        "Threat Modeling",
    ]
    
    row = 5
    for area in process_areas:
        ws[f"A{row}"] = area
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = get_border()
        
        # Current_Level
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['maturity_level'].add(cell)
        
        # Target_Level
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['maturity_level'].add(cell)
        
        # Gap (formula)
        cell = ws.cell(row=row, column=4)
        cell.value = f'=IF(OR(B{row}="",C{row}=""),"N/A",VALUE(LEFT(C{row},1))-VALUE(LEFT(B{row},1)))'
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Evidence, Action_Plan
        for col in [5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Target_Date
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        row += 1

    # Additional empty rows 15-24 (to meet minimum 20 FFFFCC data rows)
    for row in range(15, 25):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = get_border()
            if col == 4:
                cell.value = f'=IF(OR(B{row}="",C{row}=""),"N/A",VALUE(LEFT(C{row},1))-VALUE(LEFT(B{row},1)))'
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        validations['maturity_level'].add(ws.cell(row=row, column=2))
        validations['maturity_level'].add(ws.cell(row=row, column=3))
        ws.cell(row=row, column=7).number_format = 'DD.MM.YYYY'

    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Conditional formatting for gaps
    ws.conditional_formatting.add(
        'D5:D24',
        CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'D5:D24',
        CellIsRule(operator='equal', formula=['0'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    # Maturity Level Reference (row 26)
    ws.merge_cells("A26:G26")
    ws["A16"] = "MATURITY LEVEL REFERENCE"
    ws["A16"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A16"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    maturity_ref = [
        ("1 - Initial", "Ad hoc, reactive, inconsistent processes"),
        ("2 - Managed", "Documented processes, some repeatability"),
        ("3 - Defined", "Standardised processes, organisation-wide adoption"),
        ("4 - Quantitatively Managed", "Measured and controlled processes"),
        ("5 - Optimizing", "Continuous improvement, industry-leading practices"),
    ]

    row = 27
    for level, desc in maturity_ref:
        ws[f"A{row}"] = level
        ws[f"B{row}"] = desc
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws.merge_cells(f"B{row}:G{row}")
        ws[f"B{row}"].font = Font(name="Calibri", size=10)
        row += 1

    # Summary dashboard (row 33)
    ws.merge_cells("A33:D33")
    ws["A33"] = "MATURITY SUMMARY"
    ws["A33"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A33"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary = [
        ("Avg Current Maturity:", '=AVERAGE(VALUE(LEFT(B5:B24,1)))'),
        ("Avg Target Maturity:", '=AVERAGE(VALUE(LEFT(C5:C24,1)))'),
        ("Total Gap (levels):", '=SUM(D5:D24)'),
        ("Areas at Target:", '=COUNTIF(D5:D24,0)'),
    ]
    
    row = 34
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].number_format = '0.0'
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 14
    
    ws.freeze_panes = "A4"


def create_action_items(ws, styles, validations):
    """Sheet 10: Action Items - Track capability improvements."""
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "ACTION ITEMS & CAPABILITY IMPROVEMENTS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Track actions to address gaps identified in collection coverage, analysis capabilities, and process maturity."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers
    headers = [
        "Action ID", "Related Sheet", "Issue Description", "Priority", "Assigned To",
        "Due Date", "Status", "Progress Percentage", "Completion Date",
        "Outcome", "Lessons Learned", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_act = [
        "ACT-572-EXX", "Intelligence Requirements",
        "PIR coverage gap: no intelligence source for ICS/SCADA sector threats",
        "High", "J. Smith", "30.06.2026", "Open", "", "", "", "", "",
    ]
    for col_idx, val in enumerate(sample_act, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Action_ID
        cell = ws.cell(row=row, column=1)
        cell.fill = _grey_d
        cell.border = get_border()
        
        # Related_Sheet, Issue_Description, Assigned_To, Outcome, Lessons_Learned, Notes
        for col in [2, 3, 5, 10, 11, 12]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Priority
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['priority'].add(cell)
        
        # Due_Date, Completion_Date
        for col in [6, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = 'DD.MM.YYYY'
        
        # Status
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['action_status'].add(cell)
        
        # Progress_Percentage
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0%'
    
    # Apply only the validations used on this sheet (avoid shared-ref contamination)
    for key in ['priority', 'action_status']:
        ws.add_data_validation(validations[key])

    # Conditional formatting (skip sample row)
    ws.conditional_formatting.add(
        'D6:D55',
        CellIsRule(operator='equal', formula=['"Critical"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'G6:G55',
        CellIsRule(operator='equal', formula=['"Closed"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'G6:G55',
        CellIsRule(operator='equal', formula=['"Blocked"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

    # Summary section (row 57 — after sample + 50 data rows)
    ws.merge_cells("A57:D57")
    ws["A57"] = "ACTION ITEMS SUMMARY"
    ws["A57"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A57"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    summary = [
        ("Total Actions:", '=COUNTA(C6:C55)'),
        ("Critical Priority:", '=COUNTIF(D6:D55,"Critical")'),
        ("High Priority:", '=COUNTIF(D6:D55,"High")'),
        ("Open:", '=COUNTIF(G6:G55,"Open")'),
        ("In Progress:", '=COUNTIF(G6:G55,"In_Progress")'),
        ("Blocked:", '=COUNTIF(G6:G55,"Blocked")'),
        ("Resolved:", '=COUNTIF(G6:G55,"Resolved")'),
        ("Closed:", '=COUNTIF(G6:G55,"Closed")'),
    ]

    row = 58
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 35
    ws.column_dimensions['L'].width = 30
    
    ws.freeze_panes = "A4"


def create_metadata(ws, styles):
    """Sheet 11: Metadata - Workbook generation information."""
    
    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "WORKBOOK METADATA"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    generation_date = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    
    metadata = [
        ("Document ID:", "ISMS-IMP-A.5.7.2"),
        ("Document Title:", "Intelligence Collection & Analysis Assessment"),
        ("ISO 27001:2022 Control:", "A.5.7 (Threat Intelligence)"),
        ("Version:", "1.0"),
        ("Generated Date:", generation_date),
        ("Generator Script:", "generate_a57_2_collection_analysis.py"),
        ("Python Version:", "3.x"),
        ("Required Library:", "openpyxl"),
        ("", ""),
        ("Sheet Count:", "14"),
        ("Sheet 1:", "Instructions & Legend"),
        ("Sheet 2:", "Intelligence_Requirements (50 PIR/SIR/IR entries)"),
        ("Sheet 3:", "Collection_Sources (30 source entries)"),
        ("Sheet 4:", "Raw_Intelligence_Log (100 raw intel entries)"),
        ("Sheet 5:", "Intelligence_Production (60 products)"),
        ("Sheet 6:", "MITRE_Mapping (100 techniques)"),
        ("Sheet 7:", "Quality_Metrics (20 KPIs)"),
        ("Sheet 8:", "Vulnerability_Linked_Threats (50 VTL records) **CRITICAL**"),
        ("Sheet 9:", "Process_Maturity (CMM assessment)"),
        ("Sheet 10:", "Action_Items (50 actions)"),
        ("Sheet 11:", "Analysis_Tools (20 tools)"),
        ("Sheet 12:", "Threat_Actor_Profiles (30 actors)"),
        ("Sheet 13:", "Campaign_Tracking (20 campaigns)"),
        ("Sheet 14:", "Metadata (this sheet)"),
        ("", ""),
        ("VTL Integration:", "Sheet 8 implements VulnerabilityThreatLink schema"),
        ("Control 8.8 Link:", "Bidirectional data flow for vulnerability prioritization"),
        ("Emergency Patching:", "Active_Exploitation triggers emergency patch in 8.8"),
        ("", ""),
        ("Review Cycle:", "Quarterly"),
        ("Policy Reference:", "ISMS-POL-A.5.7 (All Sections)"),
        ("Related Workbooks:", "ISMS-IMP-A.5.7.1 (Sources), ISMS-IMP-A.5.7.3 (Integration)"),
        ("", ""),
        ("", "This workbook generates evidence, not theater. Use it well."),
    ]
    
    row = 3
    for label, value in metadata:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        
        if label in ["Document ID:", "Document Title:", "VTL Integration:", "Emergency Patching:"]:
            ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        else:
            ws[f"A{row}"].font = Font(name="Calibri", size=10)
            ws[f"B{row}"].font = Font(name="Calibri", size=10)
        
        # Highlight VTL-related rows
        if "VTL" in label or "8.8" in label or "Emergency" in label:
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
        
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


# ============================================================================
# MAIN EXECUTION
# ============================================================================


def create_analysis_tools(ws, styles, validations):
    """
    Sheet 11: Analysis_Tools - NEW v1.0
    Document threat intelligence analysis tools and platforms.
    """
    
    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "ANALYSIS TOOLS & PLATFORMS (V1.0)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Document all threat intelligence analysis tools. Track CVSS support capability for v1.0 integration."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers (16 columns)
    headers = [
        "Tool ID", "Tool Name", "Tool Category", "Vendor", "License Type",
        "Primary Users", "Use Cases", "Integration Status", "Data Sources",
        "CVSS Support", "Last Updated", "Version", "Training Required",
        "Training Status", "Cost Annual", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_tool = [
        "TOOL-001", "MISP", "OSINT Platform", "CIRCL.lu", "Open Source",
        "TI Team", "IOC Sharing / Feed Management", "Integrated",
        "OSINT / Commercial Feeds", "Yes", "25.02.2026", "2.4.180", "Yes",
        "Complete", "", "",
    ]
    for col_idx, val in enumerate(sample_tool, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Tool_ID (col 1)
        cell = ws.cell(row=row, column=1)
        cell.fill = _grey_d
        cell.border = get_border()

        # Tool_Name (col 2)
        cell = ws.cell(row=row, column=2)
        cell.fill = _yell
        cell.border = get_border()

        # Tool_Category (col 3)
        cell = ws.cell(row=row, column=3)
        cell.fill = _yell
        cell.border = get_border()
        validations['tool_category'].add(cell)

        # Vendor, Primary_Users, Use_Cases (cols 4-7)
        for col in [4, 5, 6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = get_border()

        # License_Type (col 5)
        cell = ws.cell(row=row, column=5)
        validations['license_type'].add(cell)

        # Integration_Status (col 8)
        cell = ws.cell(row=row, column=8)
        cell.fill = _yell
        cell.border = get_border()
        validations['integration_status'].add(cell)

        # Data_Sources (col 9)
        cell = ws.cell(row=row, column=9)
        cell.fill = _yell
        cell.border = get_border()

        # CVSS_Support (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = _yell
        cell.border = get_border()
        validations['yes_no_na'].add(cell)

        # Last_Updated (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Version (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = _yell
        cell.border = get_border()

        # Training_Required (col 13)
        cell = ws.cell(row=row, column=13)
        cell.fill = _yell
        cell.border = get_border()
        validations['yes_no'].add(cell)

        # Training_Status (col 14)
        cell = ws.cell(row=row, column=14)
        cell.fill = _yell
        cell.border = get_border()

        # Cost_Annual (col 15)
        cell = ws.cell(row=row, column=15)
        cell.fill = _yell
        cell.border = get_border()
        cell.number_format = '#,##0.00'

        # Notes (col 16)
        cell = ws.cell(row=row, column=16)
        cell.fill = _yell
        cell.border = get_border()

    # Conditional formatting (skip sample row)
    ws.conditional_formatting.add(
        'H6:H55',
        CellIsRule(operator='equal', formula=['"Deprecated"'],
                   fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"))
    )
    
    # Column widths
    widths = {'A': 12, 'B': 25, 'C': 18, 'D': 20, 'E': 16, 'F': 20, 'G': 35, 'H': 18,
              'I': 25, 'J': 14, 'K': 12, 'L': 12, 'M': 16, 'N': 16, 'O': 12, 'P': 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"


def create_threat_actor_profiles(ws, styles, validations):
    """
    Sheet 12: Threat_Actor_Profiles - NEW v1.0
    Maintain profiles of known threat actors targeting organisation or sector.
    """
    
    # Title
    ws.merge_cells("A1:T1")
    ws["A1"] = "THREAT ACTOR PROFILES (V1.0)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:T2")
    ws["A2"] = "Document known threat actors. CVEs_Exploited column integrates with Sheet 8 VTL for tracking."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers (20 columns)
    headers = [
        "Actor ID", "Actor Name", "Actor Aliases", "Actor Type",
        "Attribution Confidence", "Country of Origin", "First Observed",
        "Last Activity", "Targeting Our Sector", "Targeting Our Org",
        "Primary Motivation", "Sophistication Level", "Primary TTPs",
        "Common Malware", "Infrastructure Notes", "CVEs Exploited",
        "VTL Records Count", "Related Campaigns", "Last Update Date", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Row 5: F2F2F2 grey sample row (format guide — do not edit)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    _yell = PatternFill("solid", fgColor="FFFFCC")
    _grey_d = PatternFill("solid", fgColor="D9D9D9")
    sample_actor = [
        "ACTOR-001", "APT29 (Cozy Bear)", "Midnight Blizzard / The Dukes",
        "Nation-State", "High", "Russia", "2008", "25.02.2026",
        "Yes", "Confirmed", "Espionage", "Advanced",
        "Spearphishing / Living off the Land", "LOSTKEYS", "TLS beaconing via CDN",
        "CVE-2024-21413", "3", "CAM-001", "25.02.2026", "",
    ]
    for col_idx, val in enumerate(sample_actor, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = _grey
        cell.border = get_border()
        if val:
            cell.value = val

    # Data rows (50 rows: rows 6-55)
    for row in range(6, 56):
        # Actor_ID (col 1)
        cell = ws.cell(row=row, column=1)
        cell.fill = _grey_d
        cell.border = get_border()
        
        # Actor_Name, Actor_Aliases (cols 2-3)
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Actor_Type (col 4)
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['threat_actor_type'].add(cell)
        
        # Attribution_Confidence (col 5)
        cell = ws.cell(row=row, column=5)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['confidence_level'].add(cell)
        
        # Country_of_Origin (col 6)
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # First_Observed, Last_Activity (cols 7-8)
        for col in [7, 8]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = 'DD.MM.YYYY'
        
        # Targeting_Our_Sector (col 9)
        cell = ws.cell(row=row, column=9)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['yes_no_unknown'].add(cell)
        
        # Targeting_Our_Org (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['targeting_status'].add(cell)
        
        # Primary_Motivation (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['motivation'].add(cell)
        
        # Sophistication_Level (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['sophistication'].add(cell)
        
        # Primary_TTPs, Common_Malware, Infrastructure_Notes (cols 13-15)
        for col in [13, 14, 15]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # CVEs_Exploited (col 16) - INTEGRATION with Sheet 8
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # VTL_Records_Count (col 17)
        cell = ws.cell(row=row, column=17)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0'
        
        # Related_Campaigns (col 18)
        cell = ws.cell(row=row, column=18)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Last_Update_Date (col 19)
        cell = ws.cell(row=row, column=19)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # Notes (col 20)
        cell = ws.cell(row=row, column=20)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
    
    # Conditional formatting
    ws.conditional_formatting.add(
        'J6:J55',
        CellIsRule(operator='equal', formula=['"Confirmed"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'J6:J55',
        CellIsRule(operator='equal', formula=['"Suspected"'],
                   fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    
    # Column widths
    widths = {'A': 12, 'B': 25, 'C': 25, 'D': 18, 'E': 18, 'F': 18, 'G': 12, 'H': 12,
              'I': 18, 'J': 16, 'K': 18, 'L': 16, 'M': 30, 'N': 25, 'O': 35, 'P': 30,
              'Q': 14, 'R': 25, 'S': 12, 'T': 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"


def create_campaign_tracking(ws, styles, validations):
    """
    Sheet 13: Campaign_Tracking - NEW v1.0
    Track and analyse threat campaigns with VLOOKUP integration to Sheet 12.
    """
    
    # Title
    ws.merge_cells("A1:X1")
    ws["A1"] = "CAMPAIGN TRACKING (V1.0 - VLOOKUP + CVSS INTEGRATION)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:X2")
    ws["A2"] = "Track threat campaigns. Actor_Name auto-populated via VLOOKUP from Sheet 12. CVEs_CVSS_Max references Sheet 8."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column headers (24 columns)
    headers = [
        "Campaign ID", "Campaign Name", "Actor ID", "Actor Name",
        "Campaign Start Date", "Campaign End Date", "Campaign Status",
        "Target Sectors", "Target Geographies", "Our Sector Targeted",
        "Our Org Targeted", "Primary Objective", "Attack Vectors",
        "TTPs Used", "CVEs Exploited", "CVEs CVSS Max",
        "IOCs Count", "VTL Records Created", "Incidents Our Org",
        "Intelligence Sources", "Threat Level", "Monitoring Status",
        "Last Update Date", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    
    # Sample row (row 5 — F2F2F2 grey, format guide)
    _grey = PatternFill("solid", fgColor="F2F2F2")
    sample_cam = [
        "CAM-2026-EXX", "EMEA Financial Sector Spearphishing", "ACTOR-001",
        "APT29 (Cozy Bear)", "01.01.2026", "31.03.2026", "Active",
        "Finance / Banking", "EMEA", "Yes", "Confirmed", "Espionage",
        "Spearphishing / BEC", "T1566 / T1078 / T1071", "CVE-2024-21413",
        "7.8", "15", "8", "1", "MISP / OSINT", "High", "Active",
        "25.02.2026", "",
    ]
    for col_num, val in enumerate(sample_cam, start=1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = val
        cell.fill = _grey
        cell.border = get_border()
        cell.font = Font(name="Calibri", size=9, italic=True, color="666666")

    # Data rows (50 campaign capacity)
    for row in range(6, 56):
        # Campaign_ID (col 1)
        cell = ws.cell(row=row, column=1)
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()

        # Campaign_Name (col 2)
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Actor_ID (col 3) - dropdown references Sheet 12
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Actor_Name (col 4) - VLOOKUP from Sheet 12
        cell = ws.cell(row=row, column=4)
        formula = "=IFERROR(VLOOKUP(C" + str(row) + ",\'Threat Actor Profiles\'!A:B,2,FALSE),\"Unknown\")"
        cell.value = formula
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()

        # Campaign_Start_Date, Campaign_End_Date (cols 5-6)
        for col in [5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = 'DD.MM.YYYY'

        # Campaign_Status (col 7)
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['campaign_status'].add(cell)

        # Target_Sectors, Target_Geographies (cols 8-9)
        for col in [8, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()

        # Our_Sector_Targeted (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['yes_no_unknown'].add(cell)

        # Our_Org_Targeted (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['targeting_status'].add(cell)

        # Primary_Objective (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['campaign_objective'].add(cell)

        # Attack_Vectors, TTPs_Used, CVEs_Exploited (cols 13-15)
        for col in [13, 14, 15]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()

        # CVEs_CVSS_Max (col 16) - References Sheet 8 CVSS data
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0.0'
        cell.value = ""  # User to populate based on Sheet 8 data

        # IOCs_Count, VTL_Records_Created, Incidents_Our_Org (cols 17-19)
        for col in [17, 18, 19]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = '0'

        # Intelligence_Sources (col 20)
        cell = ws.cell(row=row, column=20)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Threat_Level (col 21)
        cell = ws.cell(row=row, column=21)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['threat_level'].add(cell)

        # Monitoring_Status (col 22)
        cell = ws.cell(row=row, column=22)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['monitoring_status'].add(cell)

        # Last_Update_Date (col 23)
        cell = ws.cell(row=row, column=23)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Notes (col 24)
        cell = ws.cell(row=row, column=24)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

    # Conditional formatting
    ws.conditional_formatting.add(
        'K6:K55',
        CellIsRule(operator='equal', formula=['"Confirmed"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'U6:U55',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                   font=Font(bold=True))
    )
    
    # Column widths
    widths = {'A': 16, 'B': 30, 'C': 12, 'D': 25, 'E': 12, 'F': 12, 'G': 14, 'H': 30,
              'I': 25, 'J': 16, 'K': 16, 'L': 20, 'M': 30, 'N': 30, 'O': 30, 'P': 12,
              'Q': 12, 'R': 14, 'S': 14, 'T': 25, 'U': 14, 'V': 16, 'W': 12, 'X': 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"




def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Intelligence Requirements — define priority intelligence requirements (PIRs) for the organisation.',
        '2. Complete Collection Sources — map sources to each intelligence requirement.',
        '3. Log raw intelligence in Raw Intelligence Log with source, date, and classification.',
        '4. Complete Intelligence Production — document analysis outputs, TTPs, and actionable findings.',
        '5. Complete MITRE Mapping — map threat intelligence to MITRE ATT&CK techniques.',
        '6. Review Quality Metrics — assess timeliness, accuracy, and completeness of intelligence.',
        '7. Complete Process Maturity — assess the maturity of collection and analysis processes.',
        '8. Maintain the Evidence Register with analysis records and dissemination logs.',
        '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A23"] = "Status Legend"
    ws["A23"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=24, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 25 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(ws, styles):
    """Summary Dashboard — A.5.7.2 Intelligence Collection & Analysis.

    Gold Standard implementation per A.8.33-34 reference pattern:
    - A1: em dash title with "— SUMMARY DASHBOARD" (GS-SD-014)
    - A2: 003366 italic, left-aligned, NO fill
    - TABLE 1: 003366 banner, 4472C4 section headers, D9D9D9 col headers, FFFFCC data
    - TABLE 2: 003366 banner, D9D9D9 col headers (NOT 003366), FFFFFF data rows
    - TABLE 3: C00000 banner, D9D9D9 col headers, FFFFFF data rows
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    def _f(h):
        return PatternFill("solid", fgColor=h)

    def _b():
        t = Side(style="thin")
        return Border(left=t, right=t, top=t, bottom=t)

    def _banner(r, v, fill, fc="FFFFFF", sz=12):
        """Render a full-width section banner (cols A:F merged)."""
        cell = ws.cell(row=r, column=1)
        cell.value = v
        cell.font = Font(name="Calibri", bold=True, color=fc, size=sz)
        cell.fill = _f(fill)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{r}:F{r}")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _b()

    def _col_hdr(r, labels):
        """Render D9D9D9 column headers."""
        for i, lbl in enumerate(labels):
            cell = ws.cell(row=r, column=1 + i)
            cell.value = lbl
            cell.font = Font(name="Calibri", bold=True, color="000000", size=10)
            cell.fill = _f("D9D9D9")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _b()

    def _dat(r, c, v, fill="FFFFCC", fc="000000", bold=False, num=False):
        cell = ws.cell(row=r, column=c)
        cell.value = v
        cell.font = Font(name="Calibri", bold=bold, color=fc)
        cell.fill = _f(fill)
        cell.alignment = Alignment(
            horizontal="center" if num else "left",
            vertical="center", wrap_text=True
        )
        cell.border = _b()
        return cell

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A3"

    # ── Row 1: Title (GS-SD-014: must contain "— SUMMARY DASHBOARD") ────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ISMS-IMP-A.5.7.2 \u2014 INTELLIGENCE COLLECTION & ANALYSIS \u2014 SUMMARY DASHBOARD"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.fill = _f("003366")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, 7):
        ws.cell(row=1, column=col).border = _b()
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (003366 italic, left-aligned, NO fill) ──────────────
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "ISO/IEC 27001:2022 | Control A.5.7 | Intelligence product status, action tracking, and collection KPIs"
    c.font = Font(name="Calibri", italic=True, color="003366", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 1 — STATUS DISTRIBUTION
    # ══════════════════════════════════════════════════════════════════════
    _banner(4, "TABLE 1: STATUS DISTRIBUTION", "003366")

    # Section A — Intelligence Production Status (col N, rows 6:55)
    _banner(5, "Section A: Intelligence Product Status ('Intelligence Production' \u2014 Column N, rows 6:55)", "4472C4", sz=10)
    _col_hdr(6, ["Status", "Count", "% of Total"])

    prod_statuses = ["Draft", "Review", "Published", "Archived"]
    for i, status in enumerate(prod_statuses):
        r = 7 + i
        _dat(r, 1, status)
        _dat(r, 2, f'=COUNTIF(\'Intelligence Production\'!N6:N55,\"{status}\")', num=True)
        _dat(r, 3, f'=IF($B$11=0,"\u2014",TEXT(B{r}/$B$11,"0.0%"))', num=True)

    # TOTAL (row 11)
    _dat(11, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(11, 2, "=SUM(B7:B10)", "D9D9D9", num=True)
    _dat(11, 3, "\u2014", "D9D9D9", num=True)

    # Section B — Action Items Status (col G, rows 6:55)
    _banner(12, "Section B: Action Items Status ('Action Items' \u2014 Column G, rows 6:55)", "4472C4", sz=10)
    _col_hdr(13, ["Status", "Count", "% of Total"])

    action_statuses = ["Open", "In_Progress", "Blocked", "Resolved", "Closed"]
    for i, status in enumerate(action_statuses):
        r = 14 + i
        _dat(r, 1, status)
        _dat(r, 2, f'=COUNTIF(\'Action Items\'!G6:G55,\"{status}\")', num=True)
        _dat(r, 3, f'=IF($B$19=0,"\u2014",TEXT(B{r}/$B$19,"0.0%"))', num=True)

    # TOTAL (row 19)
    _dat(19, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(19, 2, "=SUM(B14:B18)", "D9D9D9", num=True)
    _dat(19, 3, "\u2014", "D9D9D9", num=True)

    # Row 20: buffer
    for c in range(1, 7):
        ws.cell(row=20, column=c).fill = _f("FFFFFF")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 2 — KEY PERFORMANCE INDICATORS
    # Gold Standard: 003366 banner, D9D9D9 col headers, FFFFFF data rows
    # ══════════════════════════════════════════════════════════════════════
    _banner(21, "TABLE 2: KEY PERFORMANCE INDICATORS", "003366")
    _col_hdr(22, ["KPI Metric", "Value", "Notes"])
    ws.merge_cells("C22:F22")
    ws.cell(row=22, column=3).alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("Total Intelligence Products",
         "=COUNTA('Intelligence Production'!B6:B55)",
         "All products logged in Intelligence Production (col B = product name)"),
        ("Published Products",
         "=COUNTIF('Intelligence Production'!N6:N55,\"Published\")",
         "Finalised products ready for consumption"),
        ("Products in Draft or Review",
         "=COUNTIF('Intelligence Production'!N6:N55,\"Draft\")+COUNTIF('Intelligence Production'!N6:N55,\"Review\")",
         "Products still in progress"),
        ("Total Collection Sources",
         "=COUNTA('Collection Sources'!A6:A55)",
         "All intelligence collection sources documented"),
        ("Open Action Items",
         "=COUNTIF('Action Items'!G6:G55,\"Open\")",
         "Actions not yet started \u2014 requires attention"),
        ("Blocked Action Items",
         "=COUNTIF('Action Items'!G6:G55,\"Blocked\")",
         "Actions blocked \u2014 requires escalation"),
    ]
    for i, (metric, formula, note) in enumerate(kpis):
        r = 23 + i
        # Col A: metric label — FFFFFF fill, 000000 font, NOT bold (GS-SD-015)
        cell_a = ws.cell(row=r, column=1)
        cell_a.value = metric
        cell_a.font = Font(name="Calibri", color="000000")
        cell_a.fill = _f("FFFFFF")
        cell_a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_a.border = _b()
        # Col B: value formula — FFFFFF fill
        cell_b = ws.cell(row=r, column=2)
        cell_b.value = formula
        cell_b.font = Font(name="Calibri", color="000000")
        cell_b.fill = _f("FFFFFF")
        cell_b.alignment = Alignment(horizontal="center", vertical="center")
        cell_b.border = _b()
        # Col C:F merged: notes — FFFFCC fill
        cell_c = ws.cell(row=r, column=3)
        cell_c.value = note
        cell_c.font = Font(name="Calibri", color="000000")
        cell_c.fill = _f("FFFFFF")
        cell_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_c.border = _b()
        ws.merge_cells(f"C{r}:F{r}")

    # Row 29: buffer
    for c in range(1, 7):
        ws.cell(row=29, column=c).fill = _f("FFFFFF")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 3 — CRITICAL FINDINGS
    # Gold Standard: C00000 banner, D9D9D9 col headers, FFFFFF data rows
    # ══════════════════════════════════════════════════════════════════════
    _banner(30, "TABLE 3: CRITICAL FINDINGS", "C00000")
    _col_hdr(31, ["Finding", "Count", "Action Required"])
    ws.merge_cells("C31:F31")
    ws.cell(row=31, column=3).alignment = Alignment(horizontal="center", vertical="center")

    critical = [
        ("Blocked Action Items",
         "=COUNTIF('Action Items'!G6:G55,\"Blocked\")",
         "Identify blockers and escalate to resolve immediately"),
        ("Intelligence Products in Draft",
         "=COUNTIF('Intelligence Production'!N6:N55,\"Draft\")",
         "Review stale drafts and either publish or archive"),
        ("Archived Products (no longer active)",
         "=COUNTIF('Intelligence Production'!N6:N55,\"Archived\")",
         "Confirm archived products are intentionally retired"),
    ]
    for i, (finding, formula, action) in enumerate(critical):
        r = 32 + i
        c1 = ws.cell(row=r, column=1)
        c1.value = finding
        c1.font = Font(name="Calibri", bold=True, color="000000")
        c1.fill = _f("FFFFCC")
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c1.border = _b()

        c2 = ws.cell(row=r, column=2)
        c2.value = formula
        c2.font = Font(name="Calibri", color="000000")
        c2.fill = _f("FFFFCC")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _b()

        c3 = ws.cell(row=r, column=3)
        c3.value = action
        c3.font = Font(name="Calibri", color="000000")
        c3.fill = _f("FFFFCC")
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3.border = _b()
        ws.merge_cells(f"C{r}:F{r}")


def create_evidence_register(ws):
    """Create Evidence Register worksheet (Gold Standard — 100 data rows, navy headers)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border

    # Row 2: SUBTITLE
    ws.merge_cells('A2:H2')
    ws['A2'] = f'{DOCUMENT_ID} | {CONTROL_REF}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    # Row 3: intentionally empty (visual separator)

    # Row 4: COLUMN HEADERS with 003366 fill
    headers = [
        'Evidence ID', 'Assessment Area', 'Evidence Type', 'Description',
        'Location / Path', 'Date Collected', 'Collected By', 'Verification Status',
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Data validations
    ev_type_dv = DataValidation(
        type='list',
        formula1='"Config File,Screenshot,Report,Log File,Test Result,Policy Document,Contract,Diagram,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type='list',
        formula1='"\u2705 Verified,Pending,\u274c Not Verified"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # Row 5: SAMPLE ROW (F2F2F2 grey)
    sample_data = {
        1: 'EV-001',
        2: 'Intelligence Collection Sources',
        3: 'Screenshot',
        4: 'MISP Community Feed configuration showing active subscription and API integration',
        5: '/evidence/a57/misp-feed-config-2026-01-15.png',
        6: datetime.now().strftime('%d.%m.%Y'),
        7: 'Security Manager',
        8: '\u2705 Verified',
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.border = border
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(
            horizontal='center' if col == 1 else 'left',
            vertical='center',
            wrap_text=True
        )
    ev_type_dv.add(ws['C5'])
    ver_status_dv.add(ws['H5'])

    # Rows 6-105: EMPTY DATA ROWS (FFFFCC yellow, 100 rows)
    for data_row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(
                horizontal='center' if col == 1 else 'left',
                vertical='center',
                wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f'C{data_row}'])
        ver_status_dv.add(ws[f'H{data_row}'])

    # Column widths and freeze
    for col, width in [('A', 15), ('B', 28), ('C', 22), ('D', 45), ('E', 45), ('F', 16), ('G', 22), ('H', 20)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A5'


def create_approval_sheet(ws):
    """Create Approval Sign-Off worksheet (Gold Standard)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL AND SIGN-OFF'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2: CONTROL REFERENCE subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f'{DOCUMENT_ID} | {CONTROL_REF}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells('A3:E3')
    ws['A3'] = 'ASSESSMENT SUMMARY'
    ws['A3'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # Summary fields
    summary_fields = [
        ('Document:', f'{DOCUMENT_ID} - {WORKBOOK_NAME}'),
        ('Assessment Period:', ''),
        ('Overall Compliance Rate:', ''),
        ('Assessment Status:', ''),
        ('Assessed By:', ''),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == '')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        if 'Assessment Status' in label:
            status_row_for_dv = row
        row += 1

    # Assessment Status dropdown
    status_dv = DataValidation(
        type='list',
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f'B{status_row_for_dv}')

    # B6: Overall Compliance Rate — Published products / Total products
    ws['B6'] = "=IFERROR('Summary Dashboard'!B9/'Summary Dashboard'!B11,\"\")"
    ws['B6'].number_format = '0.0%'

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ['Name:', 'Title:', 'Date:', 'Signature:', 'Comments:']:
            ws[f'A{current_row}'] = field
            ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'B{current_row}:E{current_row}')
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, 'COMPLETED BY (ASSESSOR)', '4472C4')
    row = _create_approver_section(row, 'REVIEWED BY (INFORMATION SECURITY OFFICER)', '4472C4')
    row = _create_approver_section(row, 'APPROVED BY (CISO)', '003366')

    # FINAL DECISION
    ws[f'A{row}'] = 'FINAL DECISION:'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border
    ws.merge_cells(f'B{row}:E{row}')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type='list',
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ['Next Review Date:', 'Review Responsible:', 'Special Considerations:']:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths and freeze
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.freeze_panes = 'A3'


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def main():
    """Main generation function — 15-sheet standalone A.5.7.2 workbook."""
    logger.info("")
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.7.2 - Intelligence Collection & Analysis Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.7 (Threat Intelligence) — Standalone")
    logger.info("=" * 80)
    logger.info("")

    wb = create_workbook()
    styles = _STYLES
    validations = setup_validations()

    logger.info("Generating sheets...")

    # Sheet 1: Instructions
    logger.info("  [1/15] Instructions")
    ws = wb["Instructions & Legend"]
    ws.sheet_view.showGridLines = False
    create_instructions(ws, styles)

    # Sheet 2: Intelligence_Requirements (PIR/SIR/IR per IMP spec)
    logger.info("  [2/15] Intelligence_Requirements")
    ws = wb["Intelligence Requirements"]
    ws.sheet_view.showGridLines = False
    create_intelligence_requirements(ws, styles, validations)

    # Sheet 3: Collection_Sources
    logger.info("  [3/15] Collection_Sources")
    ws = wb["Collection Sources"]
    ws.sheet_view.showGridLines = False
    create_collection_sources(ws, styles, validations)

    # Sheet 4: Raw_Intelligence_Log
    logger.info("  [4/15] Raw_Intelligence_Log")
    ws = wb["Raw Intelligence Log"]
    ws.sheet_view.showGridLines = False
    create_raw_intelligence_log(ws, styles, validations)

    # Sheet 5: Intelligence_Production
    logger.info("  [5/15] Intelligence_Production")
    ws = wb["Intelligence Production"]
    ws.sheet_view.showGridLines = False
    create_intelligence_production(ws, styles, validations)

    # Sheet 6: MITRE_Mapping
    logger.info("  [6/15] MITRE_Mapping")
    ws = wb["MITRE Mapping"]
    ws.sheet_view.showGridLines = False
    create_mitre_mapping(ws, styles, validations)

    # Sheet 7: Quality_Metrics
    logger.info("  [7/15] Quality_Metrics")
    ws = wb["Quality Metrics"]
    ws.sheet_view.showGridLines = False
    create_quality_metrics(ws, styles, validations)

    # Sheet 8: Process_Maturity
    logger.info("  [8/15] Process_Maturity")
    ws = wb["Process Maturity"]
    ws.sheet_view.showGridLines = False
    create_process_maturity(ws, styles, validations)

    # Sheet 9: Action_Items
    logger.info("  [9/15] Action_Items")
    ws = wb["Action Items"]
    ws.sheet_view.showGridLines = False
    create_action_items(ws, styles, validations)

    # Sheet 10: Analysis_Tools
    logger.info("  [10/15] Analysis_Tools")
    ws = wb["Analysis Tools"]
    ws.sheet_view.showGridLines = False
    create_analysis_tools(ws, styles, validations)

    # Sheet 11: Threat_Actor_Profiles
    logger.info("  [11/15] Threat_Actor_Profiles")
    ws = wb["Threat Actor Profiles"]
    ws.sheet_view.showGridLines = False
    create_threat_actor_profiles(ws, styles, validations)

    # Sheet 12: Campaign_Tracking
    logger.info("  [12/15] Campaign_Tracking")
    ws = wb["Campaign Tracking"]
    ws.sheet_view.showGridLines = False
    create_campaign_tracking(ws, styles, validations)

    # Sheet 13: Evidence Register
    logger.info("  [13/15] Evidence Register")
    ws = wb["Evidence Register"]
    ws.sheet_view.showGridLines = False
    create_evidence_register(ws)

    # Sheet 14: Summary Dashboard
    logger.info("  [14/15] Summary Dashboard")
    ws = wb["Summary Dashboard"]
    ws.sheet_view.showGridLines = False
    create_summary_dashboard_sheet(ws, styles)

    # Sheet 15: Approval Sign-Off
    logger.info("  [15/15] Approval Sign-Off")
    ws = wb["Approval Sign-Off"]
    ws.sheet_view.showGridLines = False
    create_approval_sheet(ws)

    # Finalise validations
    finalize_validations(wb)

    # Save workbook
    filename = f"ISMS-IMP-A.5.7.2_Collection_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    wb.save(output_path)
    logger.info("")
    logger.info("=" * 80)
    logger.info("SUCCESS!")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Generated: {filename}")
    logger.info("")
    logger.info("Workbook Contents (15 sheets — Gold Standard):")
    logger.info("  - Intelligence Requirements: 1 sample + 50 empty rows")
    logger.info("  - Collection Sources: 1 sample + 50 empty rows")
    logger.info("  - Raw Intelligence Log: 1 sample + 50 empty rows")
    logger.info("  - Intelligence Production: 1 sample + 50 empty rows")
    logger.info("  - MITRE Mapping: ATT&CK technique coverage")
    logger.info("  - Quality Metrics: KPI dashboard")
    logger.info("  - Process Maturity: CMM capability assessment")
    logger.info("  - Action Items: 1 sample + 50 empty rows")
    logger.info("  - Analysis Tools: 1 sample + 50 empty rows")
    logger.info("  - Threat Actor Profiles: 1 sample + 50 empty rows")
    logger.info("  - Campaign Tracking: 1 sample + 50 empty rows")
    logger.info("  - Evidence Register: 1 sample + 100 empty rows")
    logger.info("  - Summary Dashboard: TABLE 1/2/3")
    logger.info("  - Approval Sign-Off: Gold Standard")
    logger.info("")
    logger.info("=" * 80)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
