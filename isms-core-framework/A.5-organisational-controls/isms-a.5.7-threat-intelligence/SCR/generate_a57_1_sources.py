#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.7.1 - Threat Intelligence Sources Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.7: Threat Intelligence
Assessment Domain 1 of 5: Threat Intelligence Sources Portfolio Management

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific threat intelligence infrastructure, source portfolio,
and assessment requirements.

Key customization areas:
1. Threat intelligence source types and vendors (match your actual subscriptions)
2. CVSS accuracy thresholds and scoring criteria (adapt to your risk profile)
3. Coverage requirements (geographic regions, industry sectors, threat types)
4. Cost-benefit analysis parameters (budget allocation, ROI calculations)
5. Compliance thresholds and regulatory requirements (jurisdiction-specific)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.7 Threat Intelligence Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
the organisation's threat intelligence source portfolio against ISO 27001:2022
Control A.5.7 requirements.

**Purpose:**
Enables systematic assessment of threat intelligence sources, their quality,
coverage, and effectiveness in supporting proactive security decision-making
and vulnerability management workflows.

**Assessment Scope:**
- Commercial threat intelligence feeds (vendor subscriptions)
- Open-source intelligence (OSINT) sources
- Government/CERT advisories and bulletins
- Industry information sharing communities (ISACs, ISAOs)
- Internal threat intelligence generation capabilities
- CVSS scoring support and accuracy (integration with A.8.8)
- Geographic and sector coverage analysis
- Cost-benefit analysis and budget optimization
- SLA compliance and vendor performance
- Business continuity planning for intelligence disruption
- Integration with security operations workflows
- Audit evidence collection and compliance validation

**Generated Workbook Structure (15 Sheets):**
1. Instructions - Assessment guidance and methodology
2. Source_Inventory - Comprehensive source catalog with CVSS support tracking
3. Source_Evaluation - Quality assessment with CVSS accuracy metrics
4. Coverage_Matrix - Geographic, sector, and threat type coverage analysis
5. Cost_Analysis - Budget allocation and cost-benefit evaluation
6. Compliance_Check - Regulatory and policy compliance validation
7. Action_Items - Gap remediation and improvement tracking
8. Metadata - Assessment metadata and version control
9. Integration_Points - API/feed integration technical details
10. Update_Frequency - SLA compliance and timeliness tracking
11. Source_Contacts - Vendor escalation contacts and support channels
12. Vendor_SLAs - Service level agreement performance monitoring
13. API_Integration - API health monitoring and integration status
14. Source_Performance_Validation - Quarterly validation (AUDIT CRITICAL)
15. Business_Continuity_Plan - Contingency planning (AUDIT CRITICAL)

**Key Features:**
- CVSS 4.0/3.1 support tracking with accuracy assessment (VTL integration)
- Data validation with industry-standard dropdown lists
- Conditional formatting for compliance status visualization
- Automated gap identification for coverage deficiencies
- Cost-per-indicator calculations and ROI analysis
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with A.5.7.2 (Collection/Analysis) and A.8.8 (Vulnerability Mgmt)

**VulnerabilityThreatLink (VTL) Integration:**
This assessment directly supports the VTL workflow by tracking which threat
intelligence sources provide CVSS scoring and validating their accuracy.
High-quality CVSS data from threat intelligence enables automated emergency
patching workflows when active exploitation is detected.

**Integration:**
This assessment feeds into:
- A.5.7.4 Threat Intelligence Effectiveness Dashboard (consolidated metrics)
- A.5.7.2 Collection & Analysis Assessment (source utilization tracking)
- A.8.8 Vulnerability Management (CVSS accuracy for patch prioritization)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a57_1_sources.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a57_1_sources.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a57_1_sources.py --date 20250115

Output:
    File: ISMS_A_5_7_1_Sources_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)
    Sheets: 15 worksheets (see structure above)

Post-Generation Steps:
    1. Review and customize threat intelligence source types to match your portfolio
    2. Inventory all active threat intelligence subscriptions and feeds
    3. Complete source evaluation assessments for quality and accuracy
    4. Validate CVSS support and accuracy metrics (critical for VTL)
    5. Analyze coverage gaps (geographic, sector, threat types)
    6. Review cost-benefit analysis and budget allocation
    7. Verify SLA compliance and vendor performance
    8. Document integration points and API health status
    9. Complete quarterly source validation (audit requirement)
    10. Develop business continuity plan for intelligence disruption
    11. Obtain stakeholder approvals
    12. Feed results into A.5.7.4 Effectiveness Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.7
Assessment Domain:    1 of 5 (Threat Intelligence Sources Portfolio Management)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.7: Threat Intelligence Policy (Governance)
    - ISMS-IMP-A.5.7.1: Threat Intelligence Sources Implementation Guide
    - ISMS-IMP-A.5.7.2: Intelligence Collection & Analysis Assessment (Domain 2)
    - ISMS-IMP-A.5.7.3: Intelligence Integration & Distribution Assessment (Domain 3)
    - ISO 27002:2022 Implementation Guidance for Control A.5.7
    - MITRE ATT&CK Framework Integration Guidelines
    - CVSS Scoring Standard v4.0 / v3.1 Specifications

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
- Expanded from 8 to 15 sheets for comprehensive source management
    - Added CVSS_Support column to Source_Inventory (VTL integration)
    - Added CVSS accuracy tracking to Source_Evaluation (audit requirement)
    - Added Sheets 9-13: Vendor Management (Integration, SLAs, Contacts, API)
    - Added Sheet 14: Source_Performance_Validation (AUDIT CRITICAL)
    - Added Sheet 15: Business_Continuity_Plan (AUDIT CRITICAL)
    - Enhanced conditional formatting for CVSS compliance visualization
    - Improved integration with A.5.7.4 dashboard and A.8.8 workflows

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Threat Intelligence Standards:**
Threat intelligence quality varies significantly across sources. Establish
clear quality metrics including:
- Timeliness: Time from threat detection to intelligence delivery
- Accuracy: False positive/negative rates (especially for CVSS scoring)
- Relevance: Alignment with your threat model and asset portfolio
- Actionability: Integration with existing security workflows
- Coverage: Geographic, sector, and threat type breadth/depth

Review source quality quarterly and adjust portfolio accordingly.

**CVSS Accuracy Tracking (VTL Critical):**
For sources providing CVSS scores, validate accuracy against:
- NIST NVD authoritative scores (when available)
- Internal security team assessments
- Vendor-published errata and corrections

CVSS accuracy below 85% should trigger source review or replacement, as
inaccurate severity ratings undermine automated patching workflows.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will verify:
- Regular portfolio review (quarterly minimum)
- Coverage adequacy for your threat landscape
- Cost justification for commercial sources
- Integration with security operations
- Business continuity planning for intelligence disruption

Ensure Sheet 14 (Source_Performance_Validation) is completed quarterly.

**Data Protection:**
Assessment workbooks contain sensitive information including:
- Threat intelligence vendor contracts and pricing
- Integration technical details (API keys, endpoints)
- Security gaps and coverage deficiencies
- Vendor contact information

Handle in accordance with your organisation's data classification policies.
Typical classification: CONFIDENTIAL - Internal Use Only.

**Maintenance:**
Review and update assessment:
- Quarterly: Source performance validation (Sheet 14)
- Semi-annually: Coverage analysis and gap remediation
- Annually: Complete portfolio reassessment with cost-benefit review
- Ad-hoc: When threat landscape changes or new sources become available

**Quality Assurance:**
Have threat intelligence analysts and security operations personnel validate
assessments before using results for budget decisions or portfolio changes.
Cross-reference with A.5.7.2 (Collection/Analysis) for utilization metrics.

**Regulatory Alignment:**
Ensure threat intelligence sources align with regulatory requirements:
- Financial sector: Sector-specific threat intelligence (FS-ISAC, etc.)
- Healthcare: HIPAA-relevant threat intelligence
- Critical infrastructure: ICS/SCOT-specific intelligence feeds
- Geographic: Jurisdiction-specific government CERT feeds

Customize source types and compliance criteria to include regulatory mandates.

**VulnerabilityThreatLink (VTL) Dependency:**
Control A.5.7 (Threat Intelligence) is tightly integrated with Control A.8.8
(Vulnerability Management) through the VTL schema. High-quality threat
intelligence with accurate CVSS scoring enables:
- Automated emergency patching for actively exploited vulnerabilities
- Risk-based patch prioritization
- Threat-informed vulnerability management

Poor CVSS accuracy in threat intelligence directly degrades vulnerability
management effectiveness. Prioritize sources with validated CVSS accuracy >90%.

**Business Continuity:**
Complete Sheet 15 (Business_Continuity_Plan) to document:
- Backup/alternate sources for critical intelligence categories
- Degraded operations procedures if primary sources unavailable
- Escalation procedures for intelligence disruption
- Recovery time objectives (RTO) for intelligence restoration

Intelligence disruption can blind security operations - plan accordingly.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.7.1"
WORKBOOK_NAME = "Threat Intelligence Sources Assessment"
CONTROL_ID = "A.5.7"
CONTROL_NAME = "Threat Intelligence"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
# Standard color palette for Control A.5.7 (consistency across all 5 workbooks)
COLOR_GREEN_DARK = "006100"      # Compliant, excellent (e.g., CVSS 4.0 Full, ≥95% accuracy)
COLOR_GREEN = "00B050"            # Compliant, good (e.g., ≥90% accuracy)
COLOR_GREEN_LIGHT = "C6EFCE"     # Compliant, acceptable (e.g., CVSS 3.1 Full, ≥85%)
COLOR_GREEN_PALE = "E7F4E4"      # Acceptable lower bound (e.g., CVSS 4.0 Basic)
COLOR_YELLOW = "FFEB9C"          # Warning, review needed (e.g., 80-85% accuracy)
COLOR_YELLOW_PALE = "FFF4CC"     # Minor warning (e.g., CVSS 3.1 Basic)
COLOR_ORANGE = "FFEB9C"          # Non-compliant, action required (e.g., CVSS 2.0, <80%)
COLOR_RED = "FFC7CE"             # Critical non-compliant (e.g., no CVSS support, <80% CVSS accuracy)
COLOR_GRAY = "D9D9D9"            # Inactive, N/A
COLOR_BLUE_HEADER = "003366"     # Header background
COLOR_BLUE_SUB = "4472C4"        # Subheader background
COLOR_BLUE_SECTION = "4472C4"    # Section header background
COLOR_BLUE_PALE = "D8E4F8"       # Critical role highlighting
COLOR_YELLOW_INPUT = "FFFFCC"    # Input cell background
COLOR_GRAY_FORMULA = "D9D9D9"    # Formula cell background (read-only)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """
    Create workbook with all required sheets matching ISMS-IMP-A.5.7.1 v1.0 specification.
    Sheet tail order: data sheets → Evidence Register → Summary Dashboard → Approval Sign-Off
    """
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure — 15 sheets in correct tail order
    sheets = [
        "Instructions & Legend",             # 1
        "Source Inventory",                  # 2
        "Source Evaluation",                 # 3
        "Coverage Matrix",                   # 4
        "Cost Analysis",                     # 5
        "Compliance Check",                  # 6
        "Action Items",                      # 7
        "Metadata",                          # 8
        "Update Frequency",                  # 9 - SLA compliance tracking
        "Source Contacts",                   # 10 - Vendor escalation contacts
        "Vendor SLAs",                       # 11 - SLA performance tracking
        "Source Performance Validation",     # 12 - AUDIT CRITICAL
        "Evidence Register",                 # 13
        "Summary Dashboard",                 # 14
        "Approval Sign-Off",                 # 15
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    
    **UPDATED v1.0**: Added CVSS-specific color styles.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_BLUE_HEADER, end_color=COLOR_BLUE_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_BLUE_SUB, end_color=COLOR_BLUE_SUB, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_BLUE_SECTION, end_color=COLOR_BLUE_SECTION, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_YELLOW_INPUT, end_color=COLOR_YELLOW_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color=COLOR_GRAY_FORMULA, end_color=COLOR_GRAY_FORMULA, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        # CVSS-specific styles (NEW v1.0)
        "cvss_excellent": {
            "fill": PatternFill(start_color=COLOR_GREEN_LIGHT, end_color=COLOR_GREEN_LIGHT, fill_type="solid"),
        },
        "cvss_good": {
            "fill": PatternFill(start_color=COLOR_GREEN_PALE, end_color=COLOR_GREEN_PALE, fill_type="solid"),
        },
        "cvss_acceptable": {
            "fill": PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid"),
        },
        "cvss_warning": {
            "fill": PatternFill(start_color=COLOR_YELLOW_PALE, end_color=COLOR_YELLOW_PALE, fill_type="solid"),
        },
        "cvss_poor": {
            "fill": PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid"),
        },
        "cvss_critical": {
            "fill": PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"),
        },
        # Critical role highlighting (NEW v1.0 - Sheet 15)
        "critical_role": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_BLUE_PALE, end_color=COLOR_BLUE_PALE, fill_type="solid"),
        },
    }
    
    return styles


# ============================================================================
# SECTION 2: DATA VALIDATIONS LIBRARY
# ============================================================================



_STYLES = setup_styles()
def get_border():
    """
    Create new border object (avoid shared object warnings).
    
    Helper function for sheets 4-6 that were copied from original v1.0 script.
    """
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def create_data_validations():
    """
    Create library of reusable data validation rules.
    
    **UPDATED v1.0**: Added CVSS_Support, API health, role criticality, and validation status dropdowns.
    """
    validations = {}
    
    # Source Type
    validations['source_type'] = DataValidation(
        type="list",
        formula1='"Commercial,OSINT,Government,Internal,Vendor,Peer_Sharing"',
        allow_blank=False
    )
    
    # Source Category
    validations['source_category'] = DataValidation(
        type="list",
        formula1='"Threat_Feed,ISAC,Vendor_Advisory,Blog,Social_Media,Research_Report,Internal_Telemetry"',
        allow_blank=False
    )
    
    # Status
    validations['status'] = DataValidation(
        type="list",
        formula1='"Active,Inactive,Trial,Pending_Renewal,Cancelled"',
        allow_blank=False
    )
    
    # CVSS Support (NEW v1.0)
    validations['cvss_support'] = DataValidation(
        type="list",
        formula1='"4.0 Full,4.0 Basic,3.1 Full,3.1 Basic,2.0 Only,Proprietary,None"',
        allow_blank=False
    )
    
    # Admiralty Code - Source Reliability
    validations['admiralty_source'] = DataValidation(
        type="list",
        formula1='"A - Completely Reliable,B - Usually Reliable,C - Fairly Reliable,D - Not Usually Reliable,E - Unreliable,F - Cannot Judge"',
        allow_blank=False
    )
    
    # Admiralty Code - Information Credibility
    validations['admiralty_info'] = DataValidation(
        type="list",
        formula1='"1 - Confirmed,2 - Probably True,3 - Possibly True,4 - Doubtful,5 - Improbable,6 - Cannot Judge"',
        allow_blank=False
    )
    
    # Priority
    validations['priority'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    
    # Yes/No
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    
    # Yes/No/N_A
    validations['yes_no_na'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    
    # Compliance Status
    validations['compliance_status'] = DataValidation(
        type="list",
        formula1='"✅ Compliant,❌ Non_Compliant,⚠️ Under_Review,➖ Not_Applicable"',
        allow_blank=False
    )
    
    # Integration Type (NEW v1.0)
    validations['integration_type'] = DataValidation(
        type="list",
        formula1='"API,STIX/TAXII,RSS,Email,Manual,Webhook,Syslog"',
        allow_blank=False
    )
    
    # Integration Target Type (NEW v1.0)
    validations['integration_target'] = DataValidation(
        type="list",
        formula1='"TIP,SIEM,SOAR,Vuln_Scanner,EDR,Ticketing,Firewall,Proxy,Custom"',
        allow_blank=False
    )
    
    # Integration Status (NEW v1.0)
    validations['integration_status'] = DataValidation(
        type="list",
        formula1='"Active,Degraded,Failed,Inactive,Planned"',
        allow_blank=False
    )
    
    # CVSS In Feed (NEW v1.0)
    validations['cvss_in_feed'] = DataValidation(
        type="list",
        formula1='"Yes_4.0,Yes_3.1,Yes_Both,No"',
        allow_blank=False
    )
    
    # TLP Support (NEW v1.0)
    validations['tlp_support'] = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=False
    )
    
    # Frequency (NEW v1.0)
    validations['frequency'] = DataValidation(
        type="list",
        formula1='"Real_Time,Hourly,Every_4_Hours,Every_12_Hours,Daily,Weekly,Monthly,Quarterly,Ad_Hoc"',
        allow_blank=False
    )
    
    # SLA Status (NEW v1.0)
    validations['sla_status'] = DataValidation(
        type="list",
        formula1='"Met,Missed,Exceeded,N/A,Measuring"',
        allow_blank=False
    )
    
    # API Health Status (NEW v1.0)
    validations['api_health'] = DataValidation(
        type="list",
        formula1='"Healthy,Degraded,Failed,Maintenance"',
        allow_blank=False
    )
    
    # Rate Limit Status (NEW v1.0)
    validations['rate_limit_status'] = DataValidation(
        type="list",
        formula1='"OK,Warning,Critical"',
        allow_blank=False
    )
    
    # Validation Pass Status (NEW v1.0 - Sheet 14)
    validations['validation_pass'] = DataValidation(
        type="list",
        formula1='"Pass,Conditional_Pass,Fail"',
        allow_blank=False
    )
    
    # Action Required (NEW v1.0 - Sheet 14)
    validations['action_required'] = DataValidation(
        type="list",
        formula1='"None,Review,Improve,Deprecate"',
        allow_blank=False
    )
    
    # Admiralty Code Letter (NEW v1.0 - Sheet 14 simplified)
    validations['admiralty_letter'] = DataValidation(
        type="list",
        formula1='"A,B,C,D,E,F"',
        allow_blank=False
    )
    
    # Admiralty Code Number (NEW v1.0 - Sheet 14 simplified)
    validations['admiralty_number'] = DataValidation(
        type="list",
        formula1='"1,2,3,4,5,6"',
        allow_blank=False
    )
    
    # Test Result (NEW v1.0 - Sheet 15)
    validations['test_result'] = DataValidation(
        type="list",
        formula1='"Pass,Partial_Pass,Fail,Not_Tested"',
        allow_blank=False
    )
    
    # Role Category (NEW v1.0 - Sheet 15)
    validations['role_category'] = DataValidation(
        type="list",
        formula1='"Critical,Important,Standard"',
        allow_blank=False
    )
    
    # Employment Status (NEW v1.0 - Sheet 15)
    validations['employment_status'] = DataValidation(
        type="list",
        formula1='"Active,On_Leave,Departed,Other"',
        allow_blank=False
    )
    
    # Training Status (NEW v1.0 - Sheet 15)
    validations['training_status'] = DataValidation(
        type="list",
        formula1='"Yes,No,In_Progress"',
        allow_blank=False
    )
    
    # Access Documented (NEW v1.0 - Sheet 15)
    validations['access_documented'] = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=False
    )
    
    return validations


def apply_cvss_support_conditional_formatting(ws, col_letter, start_row, end_row):
    """
    Apply conditional formatting to CVSS_Support column (NEW v1.0 - Sheet 2).
    
    Args:
        ws: Worksheet object
        col_letter: Column letter (e.g., 'K')
        start_row: First data row (typically 5)
        end_row: Last data row (typically 100)
    """
    # 4.0 Full → Green
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"4.0 Full"'], fill=PatternFill(start_color=COLOR_GREEN_LIGHT, end_color=COLOR_GREEN_LIGHT, fill_type="solid"))
    )
    
    # 4.0 Basic → Light Green
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"4.0 Basic"'], fill=PatternFill(start_color=COLOR_GREEN_PALE, end_color=COLOR_GREEN_PALE, fill_type="solid"))
    )
    
    # 3.1 Full → Yellow
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"3.1 Full"'], fill=PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid"))
    )
    
    # 3.1 Basic → Light Yellow
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"3.1 Basic"'], fill=PatternFill(start_color=COLOR_YELLOW_PALE, end_color=COLOR_YELLOW_PALE, fill_type="solid"))
    )
    
    # 2.0 Only → Orange
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"2.0 Only"'], fill=PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid"))
    )
    
    # Proprietary or None → Red
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"Proprietary"'], fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"None"'], fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"))
    )


def apply_cvss_accuracy_conditional_formatting(ws, col_letter, start_row, end_row):
    """
    Apply conditional formatting to CVSS_Accuracy_Rate column (NEW v1.0 - Sheets 3, 14).
    
    Thresholds per ISMS-POL-A.5.7-S4 Section 4.4.3:
    - ≥95%: Dark green (excellent)
    - ≥90%: Green (meets requirement)
    - ≥85%: Yellow (warning)
    - <85%: Red (fails requirement)
    
    Args:
        ws: Worksheet object
        col_letter: Column letter (e.g., 'R')
        start_row: First data row
        end_row: Last data row
    """
    # ≥95% → Dark green (white text for contrast)
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=[0.95], 
                   fill=PatternFill(start_color=COLOR_GREEN_DARK, end_color=COLOR_GREEN_DARK, fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    
    # ≥90% → Green
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=[0.90], fill=PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid"))
    )
    
    # ≥85% → Yellow
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=[0.85], fill=PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid"))
    )
    
    # <85% → Red (bold text for emphasis)
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='lessThan', formula=[0.85], 
                   fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"),
                   font=Font(bold=True, color="9C0006"))
    )




# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS (UPDATED V1.0)
# ============================================================================

def create_instructions(ws, styles):
    """Create GS-IL-compliant Instructions & Legend sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 2 — empty

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Workbook Title", WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Instructions section
    row = 12
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Document all threat intelligence sources used by your organisation in the Source Inventory sheet (commercial feeds, ISACs, open-source, vendor advisories).",
        "2. Evaluate each source for reliability, timeliness, accuracy, and relevance in the Source Evaluation sheet using the scoring criteria provided.",
        "3. Classify sources by intelligence type (strategic, tactical, operational, technical) and threat category in the Source Classification sheet.",
        "4. Review and record subscription and licensing details for each paid source — including contract start/end dates, renewal status, and contact information.",
        "5. Use the dropdown menus for standardised ratings: source type, reliability score, and integration status.",
        "6. Flag sources scoring below acceptable thresholds or with lapsed contracts for review and potential replacement.",
        "7. Maintain the Evidence Register with subscription agreements, evaluation reports, sample feed outputs, and ISAC membership certificates.",
        "8. Obtain final sign-off from the Threat Intelligence Lead, Information Security Manager, and CISO.",
    ]
    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # Status Legend section
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("✓", "Compliant / Complete", "Requirement fully met", _green),
        ("⚠", "Partial / In Progress", "Partially met or in progress", _amber),
        ("✗", "Non-Compliant / Not Started", "Requirement not met", _red),
        ("—", "Not Applicable", "Not applicable to this assessment", None),
    ]
    row += 1
    for sym, status, desc, fill in legend_rows:
        ws.cell(row=row, column=1, value=sym).border = _border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    # Acceptable Evidence section
    row += 1
    ws[f"A{row}"] = "Acceptable Evidence (examples)"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for evidence in [
        "\u2713 Threat intelligence feed subscription agreements and service-level agreements",
        "\u2713 Source reliability and evaluation reports with scoring rationale",
        "\u2713 Sample intelligence feed outputs (sanitised — no classified data)",
        "\u2713 ISAC / ISAO membership certificates and access documentation",
        "\u2713 Source classification records and periodic review decisions",
    ]:
        ws[f"A{row}"] = evidence
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"
def create_source_inventory(ws, styles, validations):
    """
    Create Source_Inventory sheet - master list of all TI sources.
    
    **UPDATED v1.0**: Added CVSS_Support column (column K) with conditional formatting.
    Total columns expanded from 15 (O) to 16 (P).
    """
    
    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "THREAT INTELLIGENCE SOURCE INVENTORY (WITH CVSS CAPABILITY TRACKING)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Document all threat intelligence sources. NEW: Track CVSS v4.0/3.1 support for vulnerability intelligence integration with Control 8.8."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 4)
    headers = [
        "Source ID",           # A
        "Source Name",         # B
        "Source Type",         # C
        "Source Category",     # D
        "Provider",            # E
        "Contact Email",       # F
        "Contract Start",      # G
        "Contract End",        # H
        "Auto Renew",          # I
        "Status",              # J
        "CVSS Support",        # K - NEW v1.0
        "Primary Owner",       # L (was K)
        "Backup Owner",        # M (was L)
        "Last Review Date",    # N (was M)
        "Next Review Date",    # O (was N)
        "Notes",               # P (was O)
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    # Set column widths
    widths = [15, 35, 15, 20, 25, 30, 15, 15, 15, 15, 18, 25, 25, 15, 15, 50]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 5: F2F2F2 sample row (how-to-fill example)
    sample_data = {
        1: 'SRC-001', 2: 'MISP Community Feed', 3: 'OSINT', 4: 'Threat Feed',
        5: 'CIRCL Luxembourg', 6: 'info@circl.lu', 7: '01.01.2025', 8: '31.12.2025',
        9: 'No', 10: 'Active', 11: 'Full', 12: 'CISO',
        13: 'Security Manager', 14: '01.01.2025', 15: '01.04.2025',
        16: 'Open-source threat intelligence — primary OSINT feed',
    }
    for col, val in sample_data.items():
        cell = ws.cell(row=5, column=col, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # DVs and FFFFCC data rows (rows 6-55)
    dv_auto_renew = DataValidation(type="list", formula1='"Yes,No,Under Review"', allow_blank=True)
    ws.add_data_validation(dv_auto_renew)

    for row in range(6, 56):
        # FFFFCC fill + borders on all 16 columns
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

        validations['source_type'].add(f'C{row}')
        validations['source_category'].add(f'D{row}')
        dv_auto_renew.add(f'I{row}')
        validations['status'].add(f'J{row}')
        validations['cvss_support'].add(f'K{row}')

        # Next_Review_Date formula (O) - auto-calc from N
        ws.cell(row=row, column=15).value = f'=IF(N{row}<>"", N{row}+90, "")'
        ws.cell(row=row, column=15).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=7).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=8).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=14).number_format = 'DD.MM.YYYY'

    # Conditional formatting (data rows only, skip sample)
    apply_cvss_support_conditional_formatting(ws, 'K', 6, 55)
    ws.conditional_formatting.add(
        'J6:J55',
        CellIsRule(operator='equal', formula=['"Inactive"'], fill=PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'J6:J55',
        CellIsRule(operator='equal', formula=['"Cancelled"'], fill=PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'H6:H55',
        CellIsRule(operator='lessThan', formula=['TODAY()'], fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'O6:O55',
        CellIsRule(operator='lessThan', formula=['TODAY()'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 5: SHEET 3 - SOURCE_EVALUATION (UPDATED V1.0 - CVSS ACCURACY ADDED)
# ============================================================================

def create_source_evaluation(ws, styles, validations):
    """
    Create Source_Evaluation sheet - quality assessment using Admiralty Code.
    
    **UPDATED v1.0**: Added CVSS_Accuracy_Rate, CVSS_Sample_Size, CVSS_Validation_Date columns.
    """
    
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:W1")
    ws["A1"] = "THREAT INTELLIGENCE SOURCE EVALUATION (QUALITY + CVSS ACCURACY ASSESSMENT)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells("A2:W2")
    ws["A2"] = "Assess source reliability using Admiralty Code. NEW: Track CVSS accuracy rate per ISMS-POL-A.5.7-S4 Section 4.4.3 (target: ≥90%)."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Column headers (row 4)
    headers = [
        "Source ID",                    # A
        "Source Name",                  # B (formula)
        "Evaluation Date",              # C
        "Evaluator",                    # D
        "Reliability Rating",           # E
        "Reliability Justification",    # F
        "Credibility Rating",           # G
        "Credibility Justification",    # H
        "Timeliness Score",             # I
        "Timeliness Notes",             # J
        "Relevance Score",              # K
        "Relevance Notes",              # L
        "Actionability Score",          # M
        "Actionability Notes",          # N
        "Overall Quality Score",        # O (formula)
        "Quality Rating",               # P (formula)
        "False Positive Rate",          # Q
        "CVSS Accuracy Rate",           # R - NEW v1.0
        "CVSS Sample Size",             # S - NEW v1.0
        "CVSS Validation Date",         # T - NEW v1.0
        "Evidence Link",                # U (was R)
        "Recommendation",               # V (was S)
        "Next Evaluation",              # W (was T)
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    # Set column widths
    widths = [15, 30, 15, 25, 20, 35, 20, 35, 12, 30, 12, 30, 15, 30, 18, 15, 18, 15, 15, 15, 30, 18, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Row 5: F2F2F2 sample row
    sample = ["SRC-001", "MISP Community Feed", "15.01.2025", "Jane Smith",
              "B", "Reliable — consistent accuracy over 12 months", "2",
              "Independently corroborated by CERT-EU", 4,
              "Feeds updated daily, occasional 2h delays", 4,
              "Covers EU threats; limited Asia-Pacific", 4,
              "High actionability for network-based indicators", "", "",
              "Low", 0.91, 47, "10.01.2025",
              "/evidence/a57/source-eval-2025-01.pdf", "Continue", ""]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
    # Formula cells in sample row
    ws.cell(row=5, column=3).number_format = 'DD.MM.YYYY'
    ws.cell(row=5, column=15).value = '=IF(AND(I5<>"",K5<>"",M5<>""),AVERAGE(I5,K5,M5),"")'
    ws.cell(row=5, column=15).number_format = '0.00'
    ws.cell(row=5, column=16).value = '=IF(O5="","",IF(O5>=4.5,"Excellent",IF(O5>=3.5,"Good",IF(O5>=2.5,"Fair","Poor"))))'
    ws.cell(row=5, column=18).number_format = '0.0%'
    ws.cell(row=5, column=19).number_format = '0'
    ws.cell(row=5, column=20).number_format = 'DD.MM.YYYY'
    ws.cell(row=5, column=23).value = '=IF(C5<>"",C5+90,"")'
    ws.cell(row=5, column=23).number_format = 'DD.MM.YYYY'

    # Data validations (created once, cells added in loop below)
    dv_source = DataValidation(type="list", formula1="='Source Inventory'!$A$6:$A$55", allow_blank=True)
    ws.add_data_validation(dv_source)
    dv_score = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
    ws.add_data_validation(dv_score)
    dv_fp = DataValidation(type="list", formula1='"High,Medium,Low,Unknown"', allow_blank=True)
    ws.add_data_validation(dv_fp)
    dv_rec = DataValidation(type="list", formula1='"Continue,Enhance,Review,Discontinue"', allow_blank=True)
    ws.add_data_validation(dv_rec)
    ws.add_data_validation(validations['admiralty_source'])
    ws.add_data_validation(validations['admiralty_info'])

    # Rows 6-55: FFFFCC data rows
    for row in range(6, 56):
        for col in range(1, 24):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        dv_source.add(f'A{row}')
        ws.cell(row=row, column=2).value = f"=IF(A{row}=\"\",\"\",IFERROR(VLOOKUP(A{row},'Source Inventory'!$A$6:$P$55,2,FALSE),\"[Source Not Found]\"))"
        ws.cell(row=row, column=3).value = f'=IF(A{row}<>"",TODAY(),"")'
        ws.cell(row=row, column=3).number_format = 'DD.MM.YYYY'
        validations['admiralty_source'].add(f'E{row}')
        validations['admiralty_info'].add(f'G{row}')
        dv_score.add(f'I{row}')
        dv_score.add(f'K{row}')
        dv_score.add(f'M{row}')
        ws.cell(row=row, column=15).value = f'=IF(AND(I{row}<>"",K{row}<>"",M{row}<>""),AVERAGE(I{row},K{row},M{row}),"")'
        ws.cell(row=row, column=15).number_format = '0.00'
        ws.cell(row=row, column=16).value = f'=IF(O{row}="","",IF(O{row}>=4.5,"Excellent",IF(O{row}>=3.5,"Good",IF(O{row}>=2.5,"Fair","Poor"))))'
        dv_fp.add(f'Q{row}')
        ws.cell(row=row, column=18).number_format = '0.0%'
        ws.cell(row=row, column=19).number_format = '0'
        ws.cell(row=row, column=20).number_format = 'DD.MM.YYYY'
        dv_rec.add(f'V{row}')
        ws.cell(row=row, column=23).value = f'=IF(C{row}<>"",C{row}+90,"")'
        ws.cell(row=row, column=23).number_format = 'DD.MM.YYYY'

    # Conditional formatting (data rows only, skip sample row)
    ws.conditional_formatting.add(
        'P6:P55',
        CellIsRule(operator='equal', formula=['"Excellent"'], fill=PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'P6:P55',
        CellIsRule(operator='equal', formula=['"Good"'], fill=PatternFill(start_color=COLOR_GREEN_LIGHT, end_color=COLOR_GREEN_LIGHT, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'P6:P55',
        CellIsRule(operator='equal', formula=['"Fair"'], fill=PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'P6:P55',
        CellIsRule(operator='equal', formula=['"Poor"'], fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"))
    )
    apply_cvss_accuracy_conditional_formatting(ws, 'R', 6, 55)
    ws.conditional_formatting.add(
        'V6:V55',
        CellIsRule(operator='equal', formula=['"Discontinue"'], fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"))
    )

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6-8: SHEETS 4-6 - COVERAGE, COST, COMPLIANCE (UNCHANGED FROM V1.0)
# ============================================================================

def create_coverage_matrix(ws, styles, validations):
    """Create Coverage_Matrix sheet - geographic, sector, and threat type coverage."""
    
    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "THREAT INTELLIGENCE COVERAGE MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "Map sources to coverage dimensions to identify gaps. Minimum 2 sources recommended per critical category."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(wrap_text=True)
    
    row = 4
    
    # =======================================================================
    # SUB-TABLE 1: GEOGRAPHIC COVERAGE
    # =======================================================================
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "GEOGRAPHIC COVERAGE"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    geo_headers = [
        "Source_ID",
        "Source_Name",
        "Global",
        "North_America",
        "Europe",
        "Asia_Pacific",
        "Middle_East",
        "Latin_America",
        "Africa",
    ]
    
    col_widths_geo = [15, 25, 10, 15, 10, 15, 12, 15, 10]
    
    for col_num, (header, width) in enumerate(zip(geo_headers, col_widths_geo), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}{row}"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Geographic data rows (20 sources)
    row += 1
    geo_start_row = row
    for i in range(20):
        # Source_ID
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = get_border()
        
        # Source_Name formula
        ws[f"B{row}"] = "=IFERROR(VLOOKUP(A" + str(row) + ",'Source Inventory'!A:B,2,FALSE),\"\")"
        ws[f"B{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Checkboxes (C-I) - using Yes/No dropdowns
        for col in ["C", "D", "E", "F", "G", "H", "I"]:
            validations['yes_no'].add(ws[f"{col}{row}"])
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        
        row += 1
    
    geo_end_row = row - 1
    
    # Geographic summary row
    ws[f"A{row}"] = "Sources per Region:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = get_border()
    
    for col_num, col in enumerate(["C", "D", "E", "F", "G", "H", "I"], start=3):
        ws[f"{col}{row}"] = f'=COUNTIF({col}{geo_start_row}:{col}{geo_end_row},"Yes")'
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col}{row}"].border = get_border()
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
    
    row += 2
    
    # =======================================================================
    # SUB-TABLE 2: SECTOR COVERAGE
    # =======================================================================
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "SECTOR COVERAGE"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    sector_headers = [
        "Source_ID",
        "Source_Name",
        "Financial",
        "Healthcare",
        "Government",
        "Critical_Infra",
        "Technology",
        "Education",
        "Retail",
        "Manufacturing",
        "All_Sectors",
    ]
    
    col_widths_sector = [15, 25, 12, 12, 12, 15, 12, 12, 10, 15, 12]
    
    for col_num, (header, width) in enumerate(zip(sector_headers, col_widths_sector), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}{row}"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width
    
    # Sector data rows (20 sources)
    row += 1
    sector_start_row = row
    for i in range(20):
        # Source_ID
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = get_border()
        
        # Source_Name formula
        ws[f"B{row}"] = "=IFERROR(VLOOKUP(A" + str(row) + ",'Source Inventory'!A:B,2,FALSE),\"\")"
        ws[f"B{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Checkboxes (C-K) - using Yes/No dropdowns
        for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            validations['yes_no'].add(ws[f"{col}{row}"])
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        
        row += 1
    
    sector_end_row = row - 1
    
    # Sector summary row
    ws[f"A{row}"] = "Sources per Sector:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = get_border()
    
    # Note: col K (All_Sectors) excluded from summary to avoid cross-section formula range warnings
    for col in ["C", "D", "E", "F", "G", "H", "I", "J"]:
        ws[f"{col}{row}"] = f'=COUNTIF({col}{sector_start_row}:{col}{sector_end_row},"Yes")'
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col}{row}"].border = get_border()
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
    
    row += 2
    
    # =======================================================================
    # SUB-TABLE 3: THREAT TYPE COVERAGE
    # =======================================================================
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "THREAT TYPE COVERAGE"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    threat_headers = [
        "Source_ID",
        "Source_Name",
        "Malware",
        "Phishing",
        "Ransomware",
        "Data_Breach",
        "DDoS",
        "Insider",
        "Supply_Chain",
        "Zero_Day",
        "APT",
        "Vulnerabilities",
    ]
    
    col_widths_threat = [15, 25, 10, 10, 12, 12, 8, 10, 13, 10, 8, 15]
    
    for col_num, (header, width) in enumerate(zip(threat_headers, col_widths_threat), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}{row}"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        if col_num <= len(col_widths_threat):
            ws.column_dimensions[col_letter].width = width
    
    # Threat type data rows (20 sources)
    row += 1
    threat_start_row = row
    for i in range(20):
        # Source_ID
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = get_border()
        
        # Source_Name formula
        ws[f"B{row}"] = "=IFERROR(VLOOKUP(A" + str(row) + ",'Source Inventory'!A:B,2,FALSE),\"\")"
        ws[f"B{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        
        # Checkboxes (C-L) - using Yes/No dropdowns
        for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            validations['yes_no'].add(ws[f"{col}{row}"])
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        
        row += 1
    
    threat_end_row = row - 1
    
    # Threat type summary row
    ws[f"A{row}"] = "Sources per Threat Type:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = get_border()
    
    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws[f"{col}{row}"] = f'=COUNTIF({col}{threat_start_row}:{col}{threat_end_row},"Yes")'
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col}{row}"].border = get_border()
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
    
    # Apply validation
    ws.add_data_validation(validations['yes_no'])
    
    # Freeze panes
    ws.freeze_panes = "A4"
    
    row += 2
    
    # Gap Analysis section
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COVERAGE GAP INDICATORS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    
    row += 1
    ws[f"A{row}"] = "Categories with <2 sources indicate potential coverage gaps requiring attention."
    ws[f"A{row}"].font = Font(italic=True, size=9, color="003366")

# ============================================================================
# SECTION 7: SHEET 5 - COST_ANALYSIS
# ============================================================================

def create_cost_analysis(ws, styles, validations):
    """Create Cost_Analysis sheet - ROI and budget tracking."""
    
    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "THREAT INTELLIGENCE COST & ROI ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:N2")
    ws["A2"] = "Track costs, analyse ROI, and manage budget for all threat intelligence sources. Include subscription fees, implementation, and operational costs."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        "Source ID",
        "Source Name",
        "Annual Cost CHF",
        "Implementation Cost",
        "Operational Cost Annual",
        "Total Cost",
        "Alerts Per Month",
        "Cost Per Alert",
        "Actionable Alerts Pct",
        "Value Rating",
        "ROI Assessment",
        "Budget Next Year",
        "Renewal Date",
        "Notes",
    ]

    col_widths = [15, 25, 18, 18, 22, 15, 18, 18, 20, 15, 18, 18, 15, 30]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width

    # Row 5: F2F2F2 sample row
    sample_cost = ["SRC-001", "MISP Community Feed", 0, 2500, 1200, "", 450, "",
                   85, "Good", "Good_Value", 0, "31.12.2025",
                   "Free OSINT feed; implementation cost for API integration"]
    for col_idx, val in enumerate(sample_cost, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = get_border()
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=5, column=3).number_format = '"CHF "#,##0.00'
    ws.cell(row=5, column=4).number_format = '"CHF "#,##0.00'
    ws.cell(row=5, column=5).number_format = '"CHF "#,##0.00'
    ws.cell(row=5, column=6).value = f'=IF(AND(C5="",D5="",E5=""),"",SUM(C5:E5))'
    ws.cell(row=5, column=6).number_format = '"CHF "#,##0.00'
    ws.cell(row=5, column=7).number_format = '#,##0'
    ws.cell(row=5, column=8).value = f'=IF(OR(F5="",G5="",G5=0),"",F5/(G5*12))'
    ws.cell(row=5, column=8).number_format = '"CHF "#,##0.00'
    ws.cell(row=5, column=9).number_format = '0"%"'
    ws.cell(row=5, column=12).number_format = '"CHF "#,##0.00'
    ws.cell(row=5, column=13).number_format = 'DD.MM.YYYY'

    # Data validations (created once, cells added in loop)
    value_rating_dv = DataValidation(type="list", formula1='"Excellent,Good,Fair,Poor"', allow_blank=True)
    ws.add_data_validation(value_rating_dv)
    roi_dv = DataValidation(type="list", formula1='"High_Value,Good_Value,Acceptable,Reconsider,Discontinue"', allow_blank=True)
    ws.add_data_validation(roi_dv)

    # Rows 6-55: FFFFCC data rows
    for row in range(6, 56):
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = get_border()
        ws[f"B{row}"] = f"=IFERROR(VLOOKUP(A{row},'Source Inventory'!$A$6:$P$55,2,FALSE),\"\")"
        ws[f"B{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"B{row}"].border = get_border()
        for col in ["C", "D", "E"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = get_border()
            ws[f"{col}{row}"].number_format = '"CHF "#,##0.00'
        ws[f"F{row}"] = f'=IF(AND(C{row}="",D{row}="",E{row}=""),"",SUM(C{row}:E{row}))'
        ws[f"F{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        ws[f"F{row}"].number_format = '"CHF "#,##0.00'
        ws[f"F{row}"].font = Font(bold=True)
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        ws[f"G{row}"].number_format = '#,##0'
        ws[f"H{row}"] = f'=IF(OR(F{row}="",G{row}="",G{row}=0),"",F{row}/(G{row}*12))'
        ws[f"H{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        ws[f"H{row}"].number_format = '"CHF "#,##0.00'
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        ws[f"I{row}"].number_format = '0"%"'
        value_rating_dv.add(ws[f"J{row}"])
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        roi_dv.add(ws[f"K{row}"])
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()
        ws[f"L{row}"].number_format = '"CHF "#,##0.00'
        ws[f"M{row}"].fill = styles["input_cell"]["fill"]
        ws[f"M{row}"].border = get_border()
        ws[f"M{row}"].number_format = 'DD.MM.YYYY'
        ws[f"N{row}"].fill = styles["input_cell"]["fill"]
        ws[f"N{row}"].border = get_border()

    # Conditional formatting for ROI_Assessment (range-based, skip sample)
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"High_Value"'],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Good_Value"'],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Acceptable"'],
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Reconsider"'],
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Discontinue"'],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

    # Freeze panes
    ws.freeze_panes = "A4"

    # Summary section (row 57: data ends at row 55, buffer at 56)
    row = 57
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COST SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]

    row += 1
    ws[f"A{row}"] = "Total Annual Spend:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=SUM(C6:C55)'
    ws[f"B{row}"].font = Font(bold=True, size=11)
    ws[f"B{row}"].number_format = '"CHF "#,##0.00'

    row += 1
    ws[f"A{row}"] = "Total Implementation:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=SUM(D6:D55)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '"CHF "#,##0.00'

    row += 1
    ws[f"A{row}"] = "Total Operational:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=SUM(E6:E55)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '"CHF "#,##0.00'

    row += 1
    ws[f"A{row}"] = "Grand Total:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=SUM(F6:F55)'
    ws[f"B{row}"].font = Font(bold=True, size=12, color="006100")
    ws[f"B{row}"].number_format = '"CHF "#,##0.00'

    row += 1
    ws[f"A{row}"] = "Next Year Budget:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=SUM(L6:L55)'
    ws[f"B{row}"].font = Font(bold=True, size=11, color="0066CC")
    ws[f"B{row}"].number_format = '"CHF "#,##0.00'

    row += 2
    ws[f"A{row}"] = "Average Cost Per Alert:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=AVERAGE(H6:H55)'
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].number_format = '"CHF "#,##0.00'


# ============================================================================
# SECTION 8: SHEET 6 - COMPLIANCE_CHECK
# ============================================================================

def create_compliance_check(ws, styles, validations):
    """Create Compliance_Check sheet - legal/regulatory compliance verification."""
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "THREAT INTELLIGENCE COMPLIANCE VERIFICATION"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Verify compliance with GDPR, FADP, data processing agreements, and organisational policies for all TI sources handling personal data."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        "Source ID",
        "Source Name",
        "Handles PII",
        "GDPR Applicable",
        "FADP Applicable",
        "DPA In Place",
        "DPA Signed Date",
        "DPA Review Date",
        "SCC Required",
        "SCC In Place",
        "Compliance Status",
        "Notes",
    ]

    col_widths = [15, 25, 15, 18, 18, 15, 18, 18, 15, 15, 18, 35]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
        ws.column_dimensions[col_letter].width = width

    # Row 5: F2F2F2 sample row
    sample_cc = ["SRC-001", "MISP Community Feed", "No", "No", "No", "No",
                 "", "", "No", "No", "\u2705 Compliant", "OSINT feed; no PII processed"]
    for col_idx, val in enumerate(sample_cc, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = get_border()
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=5, column=7).number_format = 'DD.MM.YYYY'
    ws.cell(row=5, column=8).value = '=IF(G5="","",G5+365)'
    ws.cell(row=5, column=8).number_format = 'DD.MM.YYYY'

    # Rows 6-55: FFFFCC data rows
    for row in range(6, 56):
        # Source_ID
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = get_border()

        # Source_Name formula
        ws[f"B{row}"] = "=IFERROR(VLOOKUP(A" + str(row) + ",'Source Inventory'!A:B,2,FALSE),\"\")"
        ws[f"B{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"B{row}"].border = get_border()

        # Handles_PII dropdown
        validations['yes_no'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        ws[f"C{row}"].alignment = Alignment(horizontal="center")

        # GDPR_Applicable dropdown
        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        ws[f"D{row}"].alignment = Alignment(horizontal="center")

        # FADP_Applicable dropdown
        validations['yes_no'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = get_border()
        ws[f"E{row}"].alignment = Alignment(horizontal="center")

        # DPA_In_Place dropdown
        validations['yes_no'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = get_border()
        ws[f"F{row}"].alignment = Alignment(horizontal="center")

        # DPA_Signed_Date
        ws[f"G{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"].border = get_border()
        ws[f"G{row}"].number_format = 'DD.MM.YYYY'

        # DPA_Review_Date (auto-calculated: DPA_Signed + 365 days)
        ws[f"H{row}"] = f'=IF(G{row}="","",G{row}+365)'
        ws[f"H{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        ws[f"H{row}"].number_format = 'DD.MM.YYYY'

        # SCC_Required dropdown
        validations['yes_no'].add(ws[f"I{row}"])
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        ws[f"I{row}"].alignment = Alignment(horizontal="center")

        # SCC_In_Place dropdown
        validations['yes_no'].add(ws[f"J{row}"])
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{row}"].border = get_border()
        ws[f"J{row}"].alignment = Alignment(horizontal="center")

        # Compliance_Status dropdown
        validations['compliance_status'].add(ws[f"K{row}"])
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{row}"].border = get_border()

        # Notes
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        ws[f"L{row}"].border = get_border()

    # Data validations — fresh objects for this sheet only
    _yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(_yn_dv)
    for row in range(6, 56):
        for col in ["C", "D", "E", "F", "I", "J"]:
            _yn_dv.add(ws[f"{col}{row}"])

    _cs_dv = DataValidation(
        type="list",
        formula1='"\u2705 Compliant,\u274c Non_Compliant,\u26a0 Under_Review,\u2014 Not_Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(_cs_dv)
    for row in range(6, 56):
        _cs_dv.add(ws[f"K{row}"])

    # Conditional formatting for Compliance_Status (range-based, skip sample)
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Compliant"'],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Non_Compliant"'],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Under_Review"'],
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))
    ws.conditional_formatting.add('K6:K55', CellIsRule(operator='equal', formula=['"Not_Applicable"'],
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")))

    # DPA_Review_Date within 30 days → Orange warning
    ws.conditional_formatting.add('H6:H55', CellIsRule(
        operator='between',
        formula=['TODAY()', 'TODAY()+30'],
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ))

    # Freeze panes
    ws.freeze_panes = "A4"

    # Summary section (row 57: data ends at row 55, buffer at 56)
    row = 57
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLIANCE SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]

    row += 1
    ws[f"A{row}"] = "Sources Handling PII:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=COUNTIF(C6:C55,"Yes")'
    ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "DPAs in Place:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=COUNTIF(F6:F55,"Yes")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")

    row += 1
    ws[f"A{row}"] = "DPAs Missing:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=COUNTIFS(C6:C55,"Yes",F6:F55,"No")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")

    row += 1
    ws[f"A{row}"] = "Compliant Sources:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=COUNTIF(K6:K55,"Compliant")'
    ws[f"B{row}"].font = Font(bold=True, color="006100")

    row += 1
    ws[f"A{row}"] = "Non-Compliant:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=COUNTIF(K6:K55,"Non_Compliant")'
    ws[f"B{row}"].font = Font(bold=True, color="9C0006")

    row += 1
    ws[f"A{row}"] = "Under Review:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=COUNTIF(K6:K55,"Under_Review")'
    ws[f"B{row}"].font = Font(bold=True, color="C65911")


# ============================================================================
# SECTION 6: SHEET 14 - SOURCE_PERFORMANCE_VALIDATION (NEW V1.0 - AUDIT CRITICAL)
# ============================================================================

def create_source_performance_validation(ws, styles, validations):
    """
    Create Source_Performance_Validation sheet - Quarterly source accuracy validation.
    
    **NEW v1.0 - AUDIT CRITICAL**: Per ISMS-POL-A.5.7-S4 Section 4.4.3.
    Provides evidence of quarterly source validation required by ISO 27001 Control A.5.7.
    
    Target accuracy: ≥85% overall, ≥90% CVSS accuracy
    Minimum sample: 10 IOCs + 10 CVEs per source per quarter
    """
    
    # Title
    ws.merge_cells("A1:AO1")
    ws["A1"] = "SOURCE PERFORMANCE VALIDATION - QUARTERLY ACCURACY ASSESSMENT (AUDIT CRITICAL)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color=COLOR_BLUE_HEADER, end_color=COLOR_BLUE_HEADER, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:AO2")
    ws["A2"] = "QUARTERLY REQUIREMENT per ISMS-POL-A.5.7-S4 Section 4.4.3: Validate source accuracy. Sample ≥10 IOCs + ≥10 CVEs per source. Target: ≥85% overall, ≥90% CVSS. Evidence required for audit."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Column headers (row 4)
    headers = [
        "Validation ID",             # A
        "Source ID",                 # B
        "Source Name",               # C (formula)
        "Validation Quarter",        # D
        "Validation Date",           # E
        "Validator",                 # F
        "Validation Method",         # G
        "Total Sample Size",         # H
        "IOC Sample Size",           # I
        "IOC True Positives",        # J
        "IOC False Positives",       # K
        "IOC Accuracy",              # L (formula)
        "CVE Sample Size",           # M
        "CVE Accurate",              # N
        "CVE Inaccurate",            # O
        "CVE Accuracy",              # P (formula)
        "CVSS Accurate Count",       # Q
        "CVSS Inaccurate Count",     # R
        "CVSS Accuracy Rate",        # S (formula)
        "CVSS Accuracy Method",      # T
        "Overall Accuracy Rate",     # U (formula)
        "Admiralty Code Source",     # V
        "Admiralty Code Info",       # W
        "Admiralty Combined",        # X (formula)
        "Validation Pass",           # Y (formula)
        "Pass Criteria Met",         # Z (formula)
        "Action Required",           # AA (formula)
        "Action Notes",              # AB
        "Action Item Created",       # AC
        "Action Item ID",            # AD
        "Evidence Location",         # AE
        "Reviewed By",               # AF
        "Review Date",               # AG
        "Next Validation Date",      # AH (formula)
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header.replace("_", " ")
        cell.font = Font(name="Calibri", size=9, bold=True)
        cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    widths = [18, 15, 30, 15, 15, 25, 40, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 15, 40, 15, 12, 12, 12, 15, 30, 15, 40, 12, 15, 40, 25, 15, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 5: F2F2F2 sample row
    sample_spv = ["VAL-NNN", "SRC-001", "MISP Community Feed", "Q1 2025", "15.01.2025", "Jane Smith",
                  "Random sampling of IOCs and CVEs from 30-day feed capture",
                  25, 12, 11, 1, "", 13, 12, 1, "", 12, 1, "",
                  "Direct comparison with NVD CVSS scores", "", "A", "1", "", "", "", "",
                  "High quality source — continue monitoring", "No", "",
                  "/evidence/a57/val-q1-2025-src001.pdf", "John Doe", "30.01.2025", ""]
    for col_idx, val in enumerate(sample_spv, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=5, column=3).value = "=IFERROR(VLOOKUP(B5,'Source Inventory'!$A$6:$P$55,2,FALSE),\"[Not Found]\")"
    ws.cell(row=5, column=5).number_format = 'DD.MM.YYYY'
    ws.cell(row=5, column=12).value = '=IFERROR((J5/I5)*100,"")'
    ws.cell(row=5, column=12).number_format = '0.0"%"'
    ws.cell(row=5, column=16).value = '=IFERROR((N5/M5)*100,"")'
    ws.cell(row=5, column=16).number_format = '0.0"%"'
    ws.cell(row=5, column=19).value = '=IFERROR((Q5/M5)*100,"")'
    ws.cell(row=5, column=19).number_format = '0.0"%"'
    ws.cell(row=5, column=21).value = '=IFERROR(((J5+N5)/H5)*100,"")'
    ws.cell(row=5, column=21).number_format = '0.0"%"'
    ws.cell(row=5, column=24).value = '=IF(AND(V5<>"",W5<>""),CONCATENATE(V5,W5),"")'
    ws.cell(row=5, column=25).value = '=IF(AND(U5>=85,S5>=90),"Pass",IF(AND(U5>=80,S5>=85),"Conditional_Pass","Fail"))'
    ws.cell(row=5, column=26).value = '=IF(U5>=85,"Overall >=85% [OK]","Overall <85% [X]") & ", " & IF(S5>=90,"CVSS >=90% [OK]","CVSS <90% [X]")'
    ws.cell(row=5, column=27).value = '=IF(Y5="Pass","None",IF(Y5="Conditional_Pass","Review",IF(OR(V5="A",V5="B"),"Improve","Deprecate")))'
    ws.cell(row=5, column=33).number_format = 'DD.MM.YYYY'
    ws.cell(row=5, column=34).value = '=IF(E5<>"",E5+90,"")'
    ws.cell(row=5, column=34).number_format = 'DD.MM.YYYY'

    # Data validations (created once, cells added in loop)
    dv_source = DataValidation(type="list", formula1="='Source Inventory'!$A$6:$A$55", allow_blank=True)
    ws.add_data_validation(dv_source)
    ws.add_data_validation(validations['admiralty_letter'])
    ws.add_data_validation(validations['admiralty_number'])
    ws.add_data_validation(validations['yes_no_na'])

    # Rows 6-55: FFFFCC data rows
    for row in range(6, 56):
        for col in range(1, 35):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(name="Calibri", size=9)
        dv_source.add(f'B{row}')
        ws.cell(row=row, column=3).value = f"=IFERROR(VLOOKUP(B{row},'Source Inventory'!$A$6:$P$55,2,FALSE),\"[Not Found]\")"
        ws.cell(row=row, column=5).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=8).number_format = '0'
        ws.cell(row=row, column=9).number_format = '0'
        ws.cell(row=row, column=12).value = f'=IFERROR((J{row}/I{row})*100,"")'
        ws.cell(row=row, column=12).number_format = '0.0"%"'
        ws.cell(row=row, column=13).number_format = '0'
        ws.cell(row=row, column=16).value = f'=IFERROR((N{row}/M{row})*100,"")'
        ws.cell(row=row, column=16).number_format = '0.0"%"'
        ws.cell(row=row, column=19).value = f'=IFERROR((Q{row}/M{row})*100,"")'
        ws.cell(row=row, column=19).number_format = '0.0"%"'
        ws.cell(row=row, column=21).value = f'=IFERROR(((J{row}+N{row})/H{row})*100,"")'
        ws.cell(row=row, column=21).number_format = '0.0"%"'
        validations['admiralty_letter'].add(f'V{row}')
        validations['admiralty_number'].add(f'W{row}')
        ws.cell(row=row, column=24).value = f'=IF(AND(V{row}<>"",W{row}<>""),CONCATENATE(V{row},W{row}),"")'
        ws.cell(row=row, column=25).value = f'=IF(AND(U{row}>=85,S{row}>=90),"Pass",IF(AND(U{row}>=80,S{row}>=85),"Conditional_Pass","Fail"))'
        ws.cell(row=row, column=26).value = f'=IF(U{row}>=85,"Overall >=85% [OK]","Overall <85% [X]") & ", " & IF(S{row}>=90,"CVSS >=90% [OK]","CVSS <90% [X]")'
        ws.cell(row=row, column=27).value = f'=IF(Y{row}="Pass","None",IF(Y{row}="Conditional_Pass","Review",IF(OR(V{row}="A",V{row}="B"),"Improve","Deprecate")))'
        validations['yes_no_na'].add(f'AC{row}')
        ws.cell(row=row, column=33).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=34).value = f'=IF(E{row}<>"",E{row}+90,"")'
        ws.cell(row=row, column=34).number_format = 'DD.MM.YYYY'

    # Conditional formatting (data rows only, skip sample)
    ws.conditional_formatting.add('U6:U55', CellIsRule(operator='greaterThanOrEqual', formula=[0.95],
        fill=PatternFill(start_color=COLOR_GREEN_DARK, end_color=COLOR_GREEN_DARK, fill_type="solid"),
        font=Font(color="FFFFFF", bold=True)))
    ws.conditional_formatting.add('U6:U55', CellIsRule(operator='greaterThanOrEqual', formula=[0.90],
        fill=PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")))
    ws.conditional_formatting.add('U6:U55', CellIsRule(operator='greaterThanOrEqual', formula=[0.85],
        fill=PatternFill(start_color=COLOR_GREEN_LIGHT, end_color=COLOR_GREEN_LIGHT, fill_type="solid")))
    ws.conditional_formatting.add('U6:U55', CellIsRule(operator='greaterThanOrEqual', formula=[0.80],
        fill=PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid")))
    ws.conditional_formatting.add('U6:U55', CellIsRule(operator='lessThan', formula=[0.80],
        fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"),
        font=Font(color="9C0006", bold=True)))
    apply_cvss_accuracy_conditional_formatting(ws, 'S', 6, 55)
    ws.conditional_formatting.add('Y6:Y55', CellIsRule(operator='equal', formula=['"Pass"'],
        fill=PatternFill(start_color=COLOR_GREEN_LIGHT, end_color=COLOR_GREEN_LIGHT, fill_type="solid")))
    ws.conditional_formatting.add('Y6:Y55', CellIsRule(operator='equal', formula=['"Conditional_Pass"'],
        fill=PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")))
    ws.conditional_formatting.add('Y6:Y55', CellIsRule(operator='equal', formula=['"Fail"'],
        fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"),
        font=Font(bold=True)))
    ws.conditional_formatting.add('AA6:AA55', CellIsRule(operator='equal', formula=['"Deprecate"'],
        fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid"),
        font=Font(bold=True)))
    ws.conditional_formatting.add('AA6:AA55', CellIsRule(operator='equal', formula=['"Improve"'],
        fill=PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")))
    ws.conditional_formatting.add('AA6:AA55', CellIsRule(operator='equal', formula=['"Review"'],
        fill=PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid")))
    ws.conditional_formatting.add('AH6:AH55', CellIsRule(operator='lessThan', formula=['TODAY()'],
        fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")))
    ws.conditional_formatting.add('AH6:AH55', CellIsRule(operator='between', formula=['TODAY()', 'TODAY()+14'],
        fill=PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid")))

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SHEETS 10-12 - VENDOR MANAGEMENT
# ============================================================================


def create_update_frequency(ws, styles, validations):
    """Update Frequency sheet - SLA compliance tracking (Gold Standard layout)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:R1")
    ws["A1"] = "UPDATE FREQUENCY - SLA COMPLIANCE TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    for col in range(1, 19):
        ws.cell(row=1, column=col).border = border

    ws.merge_cells("A2:R2")
    ws["A2"] = f"ISO/IEC 27001:2022 | Control A.5.7 | Track update SLA compliance and timeliness for all threat intelligence sources"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Source ID", "Source Name", "Contractual Frequency", "Actual Avg Frequency",
               "Last Update Received", "Update Count Last 30 Days", "Expected Update Count",
               "Update Variance", "SLA Met", "SLA Met Justification", "Outage Count Last Quarter",
               "Longest Outage Duration", "Average Outage Duration", "Timeliness Score",
               "Timeliness Trend", "Last SLA Review", "Next SLA Review", "Notes"]
    col_widths = [15, 30, 20, 20, 20, 22, 22, 18, 12, 35, 22, 22, 22, 18, 18, 18, 18, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 5: F2F2F2 sample row
    sample = ["SRC-001", "MISP Community Feed", "Daily", "Daily", "15.02.2025",
              "28", "28", "0", "Yes", "All updates received within SLA window",
              "0", "N/A", "N/A", "98", "Stable", "01.01.2025", "01.04.2025",
              "Highly reliable feed — no issues this quarter"]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    dv_src = DataValidation(type="list", formula1="='Source Inventory'!$A$6:$A$55", allow_blank=True)
    ws.add_data_validation(dv_src)
    dv_sla = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(dv_sla)

    for row in range(6, 56):
        for col in range(1, 19):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        dv_src.add(f'A{row}')
        ws.cell(row=row, column=2).value = f"=IFERROR(VLOOKUP(A{row},'Source Inventory'!$A$6:$P$55,2,FALSE),\"\")"
        dv_sla.add(f'I{row}')
        ws.cell(row=row, column=5).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=16).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=17).value = f'=IF(P{row}<>"",P{row}+90,"")'
        ws.cell(row=row, column=17).number_format = 'DD.MM.YYYY'

    ws.conditional_formatting.add('I6:I55', CellIsRule(operator='equal', formula=['"No"'],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    ws.freeze_panes = "A5"


def create_source_contacts(ws, styles, validations):
    """Source Contacts sheet - vendor escalation contacts (Gold Standard layout)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:U1")
    ws["A1"] = "SOURCE CONTACTS - VENDOR ESCALATION AND SUPPORT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    for col in range(1, 22):
        ws.cell(row=1, column=col).border = border

    ws.merge_cells("A2:U2")
    ws["A2"] = f"ISO/IEC 27001:2022 | Control A.5.7 | Vendor escalation contacts and support channels for all threat intelligence sources"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Source ID", "Source Name", "Contact Type", "Contact Name", "Contact Title",
               "Contact Email", "Contact Phone", "Contact Region", "Availability", "Escalation Path",
               "Preferred Contact Method", "Language Supported", "Last Contact Date", "Last Contact Reason",
               "Response Quality", "Response Time", "Contact Status", "Replacement Contact",
               "Last Verified", "Next Verification", "Notes"]
    col_widths = [15, 30, 22, 25, 25, 30, 18, 18, 18, 30, 25, 20, 18, 30, 18, 18, 18, 25, 18, 18, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    sample = ["SRC-001", "MISP Community Feed", "Technical Support", "Jean-Marc Leroux",
              "Senior Threat Analyst", "jm.leroux@circl.lu", "+352 621 000001",
              "Europe", "Business Hours CET", "Email → Phone → Escalation Manager",
              "Email", "English, French", "10.01.2025", "API integration query",
              "Excellent", "< 4 hours", "Active", "",
              "01.01.2025", "01.07.2025", "Primary technical contact for API issues"]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    dv_src = DataValidation(type="list", formula1="='Source Inventory'!$A$6:$A$55", allow_blank=True)
    ws.add_data_validation(dv_src)
    dv_contact_type = DataValidation(type="list",
        formula1='"Technical Support,Account Manager,Emergency Contact,Billing,Executive,Data Protection Officer,Security Team"',
        allow_blank=True)
    ws.add_data_validation(dv_contact_type)

    for row in range(6, 56):
        for col in range(1, 22):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        dv_src.add(f'A{row}')
        ws.cell(row=row, column=2).value = f"=IFERROR(VLOOKUP(A{row},'Source Inventory'!$A$6:$P$55,2,FALSE),\"\")"
        dv_contact_type.add(f'C{row}')
        ws.cell(row=row, column=13).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=19).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=20).value = f'=IF(S{row}<>"",S{row}+180,"")'
        ws.cell(row=row, column=20).number_format = 'DD.MM.YYYY'

    ws.freeze_panes = "A5"


def create_vendor_slas(ws, styles, validations):
    """Vendor SLAs sheet - SLA performance tracking (Gold Standard layout)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:Z1")
    ws["A1"] = "VENDOR SLAS - PERFORMANCE VS. CONTRACTUAL COMMITMENTS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    for col in range(1, 27):
        ws.cell(row=1, column=col).border = border

    ws.merge_cells("A2:Z2")
    ws["A2"] = f"ISO/IEC 27001:2022 | Control A.5.7 | Track SLA performance vs. contractual commitments for all threat intelligence sources"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["SLA Record ID", "Source ID", "Source Name", "SLA Metric", "Contractual Target",
               "Contractual Target Numeric", "Actual Performance", "Actual Performance Numeric",
               "Performance Variance", "Measurement Period", "Measurement Start Date", "Measurement End Date",
               "SLA Status", "SLA Breach Count", "Penalty Clause", "Penalty Amount", "Penalty Applied",
               "Penalty Application Date", "Credit Received", "Escalated To Vendor", "Escalation Date",
               "Vendor Response", "Last Review Date", "Next Review Date", "Reviewer", "Notes"]
    col_widths = [18, 15, 30, 25, 22, 22, 22, 22, 20, 20, 22, 22, 15, 18, 15, 18, 15, 22, 15, 20, 18, 30, 18, 18, 25, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    sample = ["SLA-001", "SRC-001", "MISP Community Feed", "Update Frequency",
              "Daily", "1", "Daily", "1",
              "0", "Q1 2025", "01.01.2025", "31.03.2025",
              "Met", "0", "No", "", "No",
              "", "", "No", "", "",
              "01.04.2025", "01.07.2025", "Security Manager", "All SLAs met for Q1 2025"]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    dv_src = DataValidation(type="list", formula1="='Source Inventory'!$A$6:$A$55", allow_blank=True)
    ws.add_data_validation(dv_src)
    dv_sla_status = DataValidation(type="list", formula1='"Met,Missed,Exceeded,Disputed,Waived"', allow_blank=True)
    ws.add_data_validation(dv_sla_status)

    for row in range(6, 56):
        for col in range(1, 27):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        dv_src.add(f'B{row}')
        ws.cell(row=row, column=3).value = f"=IFERROR(VLOOKUP(B{row},'Source Inventory'!$A$6:$P$55,2,FALSE),\"\")"
        dv_sla_status.add(f'M{row}')
        ws.cell(row=row, column=9).value = f'=IF(AND(H{row}<>"",F{row}<>""),H{row}-F{row},"")'
        ws.cell(row=row, column=11).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=12).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=18).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=21).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=23).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=24).number_format = 'DD.MM.YYYY'

    ws.conditional_formatting.add('M6:M55', CellIsRule(operator='equal', formula=['"Missed"'],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    ws.conditional_formatting.add('M6:M55', CellIsRule(operator='equal', formula=['"Exceeded"'],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 9: SHEETS 7-8 - ACTION ITEMS AND SUMMARY DASHBOARD
# ============================================================================


def create_action_items(ws, styles, validations):
    """Sheet 7 - Action Items: gap remediation and improvement tracking (Gold Standard layout)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: title banner
    ws.merge_cells("A1:O1")
    ws["A1"] = "ACTION ITEMS - REMEDIATION AND IMPROVEMENT TASKS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    for col in range(1, 16):
        ws.cell(row=1, column=col).border = border

    # Row 2: subtitle
    ws.merge_cells("A2:O2")
    ws["A2"] = f"ISO/IEC 27001:2022 | Control A.5.7 | Track remediation tasks and improvements for all threat intelligence source issues"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4: column headers
    headers = ["Action ID", "Source ID", "Issue Type", "Issue Description", "Detected In Sheet",
               "Priority", "Assigned To", "Due Date", "Status", "Status Notes", "Resolution Date",
               "Evidence Link", "Created Date", "Created By", "Last Updated"]
    col_widths = [15, 15, 18, 40, 20, 12, 25, 15, 15, 30, 15, 30, 15, 25, 15]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 5: F2F2F2 sample row
    sample = ["ACT-NNN", "SRC-001", "Quality", "CVSS accuracy dropped below 85% threshold — source requires re-evaluation",
              "Source Evaluation", "High", "Security Analyst", "31.03.2025", "Open",
              "Vendor notified — awaiting corrected feed", "", "/evidence/act_src001_quality.pdf",
              "15.01.2025", "CISO", "15.01.2025"]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 6-55: FFFFCC data rows with DVs
    dv_issue = DataValidation(type="list",
        formula1='"Quality,Coverage Gap,Cost,Compliance,Contract,Integration,CVSS Accuracy,Continuity,Other"',
        allow_blank=True)
    ws.add_data_validation(dv_issue)

    dv_sheet_src = DataValidation(type="list",
        formula1='"Source Evaluation,Coverage Matrix,Cost Analysis,Compliance Check,Update Frequency,Source Contacts,Vendor SLAs,Source Performance Validation"',
        allow_blank=True)
    ws.add_data_validation(dv_sheet_src)

    dv_status = DataValidation(type="list", formula1='"Open,In Progress,Blocked,Resolved,Closed"', allow_blank=True)
    ws.add_data_validation(dv_status)

    for row in range(6, 56):
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_issue.add(f'C{row}')
        dv_sheet_src.add(f'E{row}')
        validations['priority'].add(f'F{row}')
        dv_status.add(f'I{row}')
        ws.cell(row=row, column=8).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=11).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=13).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=15).number_format = 'DD.MM.YYYY'

    # Conditional formatting (data rows only, skip sample)
    ws.conditional_formatting.add('F6:F55', CellIsRule(operator='equal', formula=['"Critical"'],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
    ws.conditional_formatting.add('F6:F55', CellIsRule(operator='equal', formula=['"High"'],
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))
    ws.conditional_formatting.add('I6:I55', CellIsRule(operator='equal', formula=['"Resolved"'],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
    ws.conditional_formatting.add('I6:I55', CellIsRule(operator='equal', formula=['"Blocked"'],
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))

    ws.freeze_panes = "A5"


def create_metadata(ws, styles):
    """
    Sheet 8 - Metadata - UPDATED v1.0.
    
    **UPDATED**: Now tracks 15 sheets instead of 8.
    Documents v1.0 changes: CVSS support, vendor management, audit evidence sheets.
    """
    ws.merge_cells("A1:D1")
    ws["A1"] = "WORKBOOK METADATA AND VERSION CONTROL"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    
    metadata = [
        ["Workbook_Version", "2.0"],
        ["Total_Sheets", "15"],
        ["Generation_Date", datetime.now().strftime("%d.%m.%Y %H:%M:%S")],
        ["Generator_Script", "generate_a57_1_sources.py"],
        ["Script_Version", "2.0.0"],
        ["Python_Version", "3.12+"],
        ["openpyxl_Version", "3.1+"],
        ["Related_Policy", "ISMS-POL-A.5.7"],
        ["Related_IMP_Spec", "ISMS-IMP-A.5.7.1 v1.0"],
        ["", ""],
        ["V1.0_Changes", ""],
        ["CVSS_Support_Added", "Sheet 2 - Column K"],
        ["CVSS_Accuracy_Added", "Sheet 3 - Columns R, S, T"],
        ["Vendor_Mgmt_Added", "Sheets 9-13 (Integration, SLAs, Contacts, API)"],
        ["Audit_Evidence_Added", "Sheets 14-15 (Validation, Continuity)"],
        ["Sheet_14_CRITICAL", "Source_Performance_Validation (Quarterly validation per POL Section 4.4.3)"],
        ["Sheet_15_CRITICAL", "Business_Continuity_Plan (Annual testing per POL Section 4.4.6)"],
    ]
    
    for row_idx, (key, value) in enumerate(metadata, start=4):
        ws.cell(row=row_idx, column=1).value = key
        ws.cell(row=row_idx, column=2).value = value
        if key in ["Workbook_Version", "Total_Sheets", "V1.0_Changes", "Sheet_14_CRITICAL", "Sheet_15_CRITICAL"]:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)


# ============================================================================
# SECTION 10: MAIN GENERATION FUNCTION
# ============================================================================



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Source Inventory — register all threat intelligence sources (OSINT, commercial, ISAC).',
        '2. Complete Source Evaluation — assess reliability, timeliness, and relevance of each source.',
        '3. Complete Coverage Matrix — verify sources cover all threat categories relevant to the organisation.',
        '4. Complete Cost Analysis — document subscription costs and value delivered per source.',
        '5. Complete Compliance Check — confirm sources meet data handling and legal requirements.',
        '6. Review Source Performance Validation — track source accuracy and false positive rates.',
        '7. Maintain the Evidence Register with source agreements and evaluation records.',
        '8. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A22"] = "Status Legend"
    ws["A22"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=23, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 24 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(ws, styles):
    """Summary Dashboard — A.5.7.1 Threat Intelligence Sources Assessment.

    Gold Standard implementation per A.8.33-34 reference pattern:
    - A1: em dash title, 003366 fill, white bold 14pt, no explicit alignment (merged center)
    - A2: subtitle, 003366 italic, left-aligned, NO fill
    - TABLE 1: 003366 banner, 4472C4 section headers, D9D9D9 col headers, FFFFCC data
    - TABLE 2: 003366 banner, D9D9D9 col headers (NOT 003366), FFFFFF data rows
    - TABLE 3: C00000 banner, D9D9D9 col headers, FFFFFF data rows
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    def _f(h):
        return PatternFill("solid", fgColor=h)

    def _b():
        t = Side(style="thin")
        return Border(left=t, right=t, top=t, bottom=t)

    def _banner(r, v, fill, fc="FFFFFF", sz=12, merge_to="F"):
        """Render a full-width section banner (cols A:F merged)."""
        cell = ws.cell(row=r, column=1)
        cell.value = v
        cell.font = Font(name="Calibri", bold=True, color=fc, size=sz)
        cell.fill = _f(fill)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{r}:{merge_to}{r}")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _b()

    def _col_hdr(r, labels):
        """Render D9D9D9 column headers for TABLE 1 (3 cols) or TABLE 2/3."""
        for i, lbl in enumerate(labels):
            cell = ws.cell(row=r, column=1 + i)
            cell.value = lbl
            cell.font = Font(name="Calibri", bold=True, color="000000", size=10)
            cell.fill = _f("D9D9D9")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _b()

    def _dat(r, c, v, fill="FFFFCC", fc="000000", bold=False, num=False):
        cell = ws.cell(row=r, column=c)
        cell.value = v
        cell.font = Font(name="Calibri", bold=bold, color=fc)
        cell.fill = _f(fill)
        cell.alignment = Alignment(
            horizontal="center" if num else "left",
            vertical="center", wrap_text=True
        )
        cell.border = _b()
        return cell

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A3"

    # ── Row 1: Title (GS-SD-014: must contain "— SUMMARY DASHBOARD") ────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ISMS-IMP-A.5.7.1 \u2014 THREAT INTELLIGENCE SOURCES ASSESSMENT \u2014 SUMMARY DASHBOARD"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.fill = _f("003366")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, 7):
        ws.cell(row=1, column=col).border = _b()
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (003366 italic, left-aligned, NO fill) ──────────────
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "ISO/IEC 27001:2022 | Control A.5.7 | Source health, compliance status, and open actions"
    c.font = Font(name="Calibri", italic=True, color="003366", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    # No fill on A2 (Gold Standard)

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 1 — STATUS DISTRIBUTION
    # ══════════════════════════════════════════════════════════════════════
    _banner(4, "TABLE 1: STATUS DISTRIBUTION", "003366")

    # Section A — Source Inventory Status (col J, rows 6:55)
    _banner(5, "Section A: Source Inventory Status ('Source Inventory' \u2014 Column J, rows 6:55)", "4472C4", sz=10)
    _col_hdr(6, ["Status", "Count", "% of Total"])

    src_statuses = ["Active", "Inactive", "Trial", "Pending_Renewal", "Cancelled"]
    for i, status in enumerate(src_statuses):
        r = 7 + i
        _dat(r, 1, status)
        _dat(r, 2, f'=COUNTIF(\'Source Inventory\'!J6:J55,\"{status}\")', num=True)
        _dat(r, 3, f'=IF($B$12=0,"\u2014",TEXT(B{r}/$B$12,"0.0%"))', num=True)

    # TOTAL (row 12)
    _dat(12, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(12, 2, "=SUM(B7:B11)", "D9D9D9", num=True)
    _dat(12, 3, "\u2014", "D9D9D9", num=True)

    # Section B — Compliance Check Status (col K, rows 6:55)
    _banner(13, "Section B: Compliance Check Status ('Compliance Check' \u2014 Column K, rows 6:55)", "4472C4", sz=10)
    _col_hdr(14, ["Status", "Count", "% of Total"])

    comp_statuses = [
        "\u2705 Compliant",
        "\u274c Non_Compliant",
        "\u26a0 Under_Review",
        "\u2014 Not_Applicable",
    ]
    for i, status in enumerate(comp_statuses):
        r = 15 + i
        _dat(r, 1, status)
        _dat(r, 2, f'=COUNTIF(\'Compliance Check\'!K6:K55,\"{status}\")', num=True)
        _dat(r, 3, f'=IF($B$19=0,"\u2014",TEXT(B{r}/$B$19,"0.0%"))', num=True)

    # TOTAL (row 19)
    _dat(19, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(19, 2, "=SUM(B15:B18)", "D9D9D9", num=True)
    _dat(19, 3, "\u2014", "D9D9D9", num=True)

    # Row 20: buffer
    for c in range(1, 7):
        ws.cell(row=20, column=c).fill = _f("FFFFFF")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 2 — KEY PERFORMANCE INDICATORS
    # Gold Standard: 003366 banner, D9D9D9 col headers, FFFFFF data rows
    # ══════════════════════════════════════════════════════════════════════
    _banner(21, "TABLE 2: KEY PERFORMANCE INDICATORS", "003366")
    _col_hdr(22, ["KPI Metric", "Value", "Notes"])
    # Merge Notes header across C:F
    ws.merge_cells("C22:F22")
    ws.cell(row=22, column=3).alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("Total Sources in Inventory",
         "=COUNTA('Source Inventory'!A6:A55)",
         "All sources tracked in Source Inventory"),
        ("Active Sources",
         "=COUNTIF('Source Inventory'!J6:J55,\"Active\")",
         "Sources with status = Active"),
        ("Compliant Sources",
         "=COUNTIF('Compliance Check'!K6:K55,\"\u2705 Compliant\")",
         "Sources meeting compliance requirements"),
        ("Non-Compliant Sources",
         "=COUNTIF('Compliance Check'!K6:K55,\"\u274c Non_Compliant\")",
         "Sources failing compliance check \u2014 requires action"),
        ("Open Action Items",
         "=COUNTIF('Action Items'!I6:I55,\"Open\")",
         "Action Items with status = Open"),
        ("Blocked Action Items",
         "=COUNTIF('Action Items'!I6:I55,\"Blocked\")",
         "Action Items blocked \u2014 requires escalation"),
    ]
    for i, (metric, formula, note) in enumerate(kpis):
        r = 23 + i
        # Col A: metric label — FFFFFF fill, 000000 font, NOT bold (GS-SD-015)
        cell_a = ws.cell(row=r, column=1)
        cell_a.value = metric
        cell_a.font = Font(name="Calibri", color="000000")
        cell_a.fill = _f("FFFFFF")
        cell_a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_a.border = _b()
        # Col B: value formula — FFFFFF fill, 000000 font
        cell_b = ws.cell(row=r, column=2)
        cell_b.value = formula
        cell_b.font = Font(name="Calibri", color="000000")
        cell_b.fill = _f("FFFFFF")
        cell_b.alignment = Alignment(horizontal="center", vertical="center")
        cell_b.border = _b()
        # Col C:F merged: notes — FFFFCC fill
        cell_c = ws.cell(row=r, column=3)
        cell_c.value = note
        cell_c.font = Font(name="Calibri", color="000000")
        cell_c.fill = _f("FFFFFF")
        cell_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_c.border = _b()
        ws.merge_cells(f"C{r}:F{r}")

    # Row 29: buffer
    for c in range(1, 7):
        ws.cell(row=29, column=c).fill = _f("FFFFFF")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 3 — CRITICAL FINDINGS
    # Gold Standard: C00000 banner, D9D9D9 col headers, FFFFFF data rows
    # ══════════════════════════════════════════════════════════════════════
    _banner(30, "TABLE 3: CRITICAL FINDINGS", "C00000")
    _col_hdr(31, ["Finding", "Count", "Action Required"])
    ws.merge_cells("C31:F31")
    ws.cell(row=31, column=3).alignment = Alignment(horizontal="center", vertical="center")

    critical = [
        ("Non-Compliant Sources",
         "=COUNTIF('Compliance Check'!K6:K55,\"\u274c Non_Compliant\")",
         "Review and remediate all non-compliant sources immediately"),
        ("Inactive or Cancelled Sources",
         "=COUNTIF('Source Inventory'!J6:J55,\"Inactive\")+COUNTIF('Source Inventory'!J6:J55,\"Cancelled\")",
         "Confirm inactive sources are intentional; archive or remove cancelled"),
        ("Blocked Action Items",
         "=COUNTIF('Action Items'!I6:I55,\"Blocked\")",
         "Escalate blocked actions \u2014 identify dependency and unblock"),
    ]
    for i, (finding, formula, action) in enumerate(critical):
        r = 32 + i
        c1 = ws.cell(row=r, column=1)
        c1.value = finding
        c1.font = Font(name="Calibri", bold=True, color="000000")
        c1.fill = _f("FFFFCC")
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c1.border = _b()

        c2 = ws.cell(row=r, column=2)
        c2.value = formula
        c2.font = Font(name="Calibri", color="000000")
        c2.fill = _f("FFFFCC")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _b()

        c3 = ws.cell(row=r, column=3)
        c3.value = action
        c3.font = Font(name="Calibri", color="000000")
        c3.fill = _f("FFFFCC")
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3.border = _b()
        ws.merge_cells(f"C{r}:F{r}")


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


# ============================================================================
# SHEET 13: EVIDENCE REGISTER (GOLD STANDARD)
# ============================================================================

def create_evidence_register(ws):
    """Create Evidence Register worksheet (Gold Standard — 100 data rows, navy headers)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border

    # Row 2: SUBTITLE
    ws.merge_cells('A2:H2')
    ws['A2'] = f'{DOCUMENT_ID} | {CONTROL_REF}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    # Row 3: intentionally empty (visual separator)

    # Row 4: COLUMN HEADERS with 003366 fill
    headers = [
        'Evidence ID', 'Assessment Area', 'Evidence Type', 'Description',
        'Location / Path', 'Date Collected', 'Collected By', 'Verification Status',
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Data validations
    ev_type_dv = DataValidation(
        type='list',
        formula1='"Config File,Screenshot,Report,Log File,Test Result,Policy Document,Contract,Diagram,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type='list',
        formula1='"\u2705 Verified,Pending,\u274c Not Verified"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # Row 5: SAMPLE ROW (F2F2F2 grey)
    sample_data = {
        1: 'EV-001',
        2: 'Threat Intelligence Sources',
        3: 'Screenshot',
        4: 'MISP Community Feed configuration showing active subscription and API integration',
        5: '/evidence/a57/misp-feed-config-2025-01-15.png',
        6: datetime.now().strftime('%d.%m.%Y'),
        7: 'Security Manager',
        8: '\u2705 Verified',
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.border = border
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(
            horizontal='center' if col == 1 else 'left',
            vertical='center',
            wrap_text=True
        )
    ev_type_dv.add(ws['C5'])
    ver_status_dv.add(ws['H5'])

    # Rows 6-105: EMPTY DATA ROWS (FFFFCC yellow, 100 rows)
    for data_row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(
                horizontal='center' if col == 1 else 'left',
                vertical='center',
                wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f'C{data_row}'])
        ver_status_dv.add(ws[f'H{data_row}'])

    # Column widths and freeze
    for col, width in [('A', 15), ('B', 28), ('C', 22), ('D', 45), ('E', 45), ('F', 16), ('G', 22), ('H', 20)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A5'


# ============================================================================
# SHEET 15: APPROVAL SIGN-OFF (GOLD STANDARD)
# ============================================================================

def create_approval_sheet(ws):
    """Create Approval Sign-Off worksheet (Gold Standard)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL AND SIGN-OFF'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2: CONTROL REFERENCE subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f'{DOCUMENT_ID} | {CONTROL_REF}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells('A3:E3')
    ws['A3'] = 'ASSESSMENT SUMMARY'
    ws['A3'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # Summary fields
    summary_fields = [
        ('Document:', f'{DOCUMENT_ID} - {WORKBOOK_NAME}'),
        ('Assessment Period:', ''),
        ('Overall Compliance Rate:', ''),
        ('Assessment Status:', ''),
        ('Assessed By:', ''),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == '')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        if 'Assessment Status' in label:
            status_row_for_dv = row
        row += 1

    # Assessment Status dropdown
    status_dv = DataValidation(
        type='list',
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f'B{status_row_for_dv}')

    # B6: Overall Compliance Rate — Compliant sources / Total (Compliance Check section)
    ws['B6'] = "=IFERROR('Summary Dashboard'!B15/'Summary Dashboard'!B19,\"\")"
    ws['B6'].number_format = '0.0%'

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ['Name:', 'Title:', 'Date:', 'Signature:', 'Comments:']:
            ws[f'A{current_row}'] = field
            ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'B{current_row}:E{current_row}')
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, 'COMPLETED BY (ASSESSOR)', '4472C4')
    row = _create_approver_section(row, 'REVIEWED BY (INFORMATION SECURITY OFFICER)', '4472C4')
    row = _create_approver_section(row, 'APPROVED BY (CISO)', '003366')

    # FINAL DECISION
    ws[f'A{row}'] = 'FINAL DECISION:'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border
    ws.merge_cells(f'B{row}:E{row}')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type='list',
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ['Next Review Date:', 'Review Responsible:', 'Special Considerations:']:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths and freeze
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.freeze_panes = 'A3'


def main():
    """
    Main function to generate complete ISMS-IMP-A.5.7.1 workbook.

    Standalone A.5.7 Threat Intelligence Sources Assessment — 15 sheets.

    Returns:
        int: 0 on success, 1 on failure
    """
    try:
        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.5.7.1 - Threat Intelligence Sources Assessment Generator")
        logger.info("Generating 15-sheet standalone workbook...")
        logger.info("=" * 80)

        # Create workbook and get styles/validations
        wb = create_workbook()
        styles = _STYLES
        validations = create_data_validations()

        # Generate all 15 sheets
        sheets = [
            ("Instructions & Legend", create_instructions, [styles]),
            ("Source Inventory", create_source_inventory, [styles, validations]),
            ("Source Evaluation", create_source_evaluation, [styles, validations]),
            ("Coverage Matrix", create_coverage_matrix, [styles, validations]),
            ("Cost Analysis", create_cost_analysis, [styles, validations]),
            ("Compliance Check", create_compliance_check, [styles, validations]),
            ("Action Items", create_action_items, [styles, validations]),
            ("Metadata", create_metadata, [styles]),
            ("Update Frequency", create_update_frequency, [styles, validations]),
            ("Source Contacts", create_source_contacts, [styles, validations]),
            ("Vendor SLAs", create_vendor_slas, [styles, validations]),
            ("Source Performance Validation", create_source_performance_validation, [styles, validations]),
            ("Evidence Register", create_evidence_register, []),
            ("Summary Dashboard", create_summary_dashboard_sheet, [styles]),
            ("Approval Sign-Off", create_approval_sheet, []),
        ]

        for sheet_name, create_func, args in sheets:
            logger.info(f"  Creating sheet: {sheet_name}...")
            ws = wb[sheet_name]
            create_func(ws, *args)

        # Finalise validations
        finalize_validations(wb)

        # Save workbook
        filename = f"ISMS-IMP-A.5.7.1_Sources_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        wb.save(output_path)
        logger.info("=" * 80)
        logger.info("SUCCESS: Workbook generated successfully!")
        logger.info(f"Filename: {filename}")
        logger.info("Sheets: 15 total")
        logger.info("   - Sheets 1-8: Core source management")
        logger.info("   - Sheets 9-12: Vendor management and audit evidence")
        logger.info("   - Sheet 13: Evidence Register")
        logger.info("   - Sheet 14: Summary Dashboard")
        logger.info("   - Sheet 15: Approval Sign-Off")
        logger.info("NEXT STEPS:")
        logger.info("   1. Populate with your organisation's data")
        logger.info("   2. Complete Sheet 12 (Source Performance Validation) quarterly")
        logger.info("=" * 80)

        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
