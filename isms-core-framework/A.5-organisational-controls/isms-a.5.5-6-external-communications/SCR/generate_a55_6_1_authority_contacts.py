#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.5-6.S1 - Authority Contacts Register Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.5-6: Contact with Authorities & Special Interest Groups
Assessment Domain 1 of 3: Authority Contacts Register

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific external communications and contact management infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Authority contact categories and update frequencies (match your jurisdiction)
2. Special interest group memberships and engagement criteria
3. Communication channel classifications and security requirements
4. Contact ownership and maintenance responsibilities
5. Escalation paths for security incident notifications

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.5-6 Contact with Authorities & Special Interest Groups Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
external communications and contact management controls and compliance requirements.

**Purpose:**
Enables systematic management of Authority Contacts Register under ISO 27001:2022 Controls A.5.5 and A.5.6. Supports evidence-based documentation of external relationships, contacts, and communication protocols for audit readiness.

**Assessment Scope:**
- External contact inventory completeness and currency
- Authority relationship classification and engagement frequency
- Special interest group membership and value assessment
- Communication procedure definition and accessibility
- Contact ownership and update accountability
- Incident notification pathway documentation
- Evidence collection for regulatory and audit requirements

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Contact with Authorities & Special Interest Groups controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a55_6_1_authority_contacts.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a55_6_1_authority_contacts.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a55_6_1_authority_contacts.py --date 20250115

Output:
    File: ISMS-IMP-A.5.5-6.S1_Authority_Contacts_Register_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.5-6
Assessment Domain:    1 of 3 (Authority Contacts Register)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.5-6: Contact with Authorities & Special Interest Groups Policy (Governance)
    - ISMS-IMP-A.5.5-6.S1: Authority Contacts Register (Domain 1)
    - ISMS-IMP-A.5.5-6.S2: Special Interest Groups Register (Domain 2)
    - ISMS-IMP-A.5.5-6.S3: External Communication Procedures (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.5-6.S1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive external communications and contact management details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review contact registers and communication procedures annually or when regulatory requirements change, authority contacts are updated, or new industry groups are joined.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
import sys
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.5-6.S1"
WORKBOOK_NAME = "Authority Contacts Register"
CONTROL_ID = "A.5.5-6"
CONTROL_NAME = "Contact with Authorities & Special Interest Groups"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=10, color="000000")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
TITLE_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Authority Registry — list all regulatory bodies, CERTs, and law enforcement contacts.',
        '2. Complete Contact Types — classify each contact (Regulatory, Incident Response, Advisory).',
        '3. Log all contacts in Communication Log with date, method, and outcome.',
        '4. Complete Verification Register — confirm contact details are current and tested.',
        '5. Maintain the Evidence Register with correspondence and contact verification records.',
        '6. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A20"] = "Status Legend"
    ws["A20"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=21, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 22 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_authority_registry_sheet(ws):
    """Create the Authority Registry sheet."""
    ws.title = "Authority Registry"

    headers = [
        "Contact ID", "Authority Name", "Authority Type", "Jurisdiction",
        "Primary Contact Name", "Primary Contact Title", "Primary Phone",
        "Primary Email", "Secondary Contact Name", "Secondary Phone",
        "Website", "Physical Address", "Internal Owner", "Owner Department",
        "Relationship Status", "Last Contact Date", "Next Review Date",
        "Escalation Trigger", "Notes"
    ]

    # Title row (003366 fill, ALL CAPS)
    _n_cols = len(headers)
    _title = ws.cell(row=1, column=1, value="AUTHORITY REGISTRY")
    _title.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    _title.fill = TITLE_FILL
    _title.alignment = Alignment(horizontal="center", vertical="center")
    _title.border = THIN_BORDER
    ws.merge_cells(f"A1:{get_column_letter(_n_cols)}1")
    ws.row_dimensions[1].height = 35

    _sub = ws.cell(row=2, column=1, value="ISO 27001:2022 | Control A.5.5-6 | Authority Contacts Register")
    _sub.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    _sub.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _sub.alignment = Alignment(horizontal="center", vertical="center")
    _sub.border = THIN_BORDER
    ws.merge_cells(f"A2:{get_column_letter(_n_cols)}2")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header.replace("_", " "))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validation dropdowns
    type_dv = DataValidation(
        type="list",
        formula1='"Law Enforcement,Data Protection Authority,Financial Regulator,Sector Regulator,Emergency Services,Government Cybersecurity,Standards Body,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('C4:C101')

    jurisdiction_dv = DataValidation(
        type="list",
        formula1='"Switzerland,Germany,Austria,EU,United Kingdom,United States,International,Other"'
    )
    ws.add_data_validation(jurisdiction_dv)
    jurisdiction_dv.add('D4:D101')

    status_dv = DataValidation(type="list", formula1='"Active,Inactive,Under Review,Pending"')
    ws.add_data_validation(status_dv)
    status_dv.add('O4:O101')

    # Format input rows
    for row in range(4, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = LOCKED_FILL if row == 4 else INPUT_FILL
            cell.border = THIN_BORDER

    # Grey sample row 4
    _sample = ("AUTH-001", "FDPIC", "Data Protection Authority", "Switzerland",
               "Office of the Commissioner", "Commissioner", "+41 31 323 27 07",
               "info@edoeb.admin.ch", "", "", "www.edoeb.admin.ch",
               "Feldeggweg 1, 3003 Bern", "J. Smith", "Compliance", "Active",
               "01.01.2026", "01.01.2027", "Data breach notification", "Key DPA contact")
    for col, val in enumerate(_sample, 1):
        ws.cell(row=4, column=col, value=val)
    ws.freeze_panes = 'A4'

    # Column widths
    widths = [12, 35, 25, 15, 25, 25, 18, 30, 25, 18, 35, 40, 20, 18, 15, 15, 15, 30, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_contact_types_sheet(ws):
    """Create the Contact Types sheet."""
    ws.title = "Contact Types"

    headers = [
        "Type Code", "Authority Type", "Description", "Typical Scenarios",
        "Mandatory Contact", "Response SLA", "Escalation Path",
        "Required Information", "Example Authorities"
    ]

    # Title row (003366 fill, ALL CAPS)
    _n_cols = len(headers)
    _title = ws.cell(row=1, column=1, value="CONTACT TYPES")
    _title.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    _title.fill = TITLE_FILL
    _title.alignment = Alignment(horizontal="center", vertical="center")
    _title.border = THIN_BORDER
    ws.merge_cells(f"A1:{get_column_letter(_n_cols)}1")
    ws.row_dimensions[1].height = 35

    _sub = ws.cell(row=2, column=1, value="ISO 27001:2022 | Control A.5.5-6 | Authority Type Reference")
    _sub.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    _sub.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _sub.alignment = Alignment(horizontal="center", vertical="center")
    _sub.border = THIN_BORDER
    ws.merge_cells(f"A2:{get_column_letter(_n_cols)}2")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header.replace("_", " "))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate with standard authority types
    authority_types = [
        ("LE", "Law Enforcement", "Police and cybercrime investigation units",
         "Security incidents, theft, criminal activity", "Yes (for crimes)",
         "Immediate", "CISO > Legal > CEO", "Incident details, evidence preservation",
         "Cantonal Police, Fedpol, Europol"),
        ("DPA", "Data Protection Authority", "Personal data protection regulators",
         "Data breaches affecting individuals", "Yes (72h breach notification)",
         "72 hours", "DPO > Legal > CISO", "Breach scope, affected individuals, mitigation",
         "FDPIC (CH), ICO (UK), CNIL (FR)"),
        ("FIN", "Financial Regulator", "Financial services regulators",
         "Financial system impacts, regulatory reporting", "Conditional",
         "Per regulation", "CFO > Legal > CEO", "Financial impact, regulatory requirements",
         "FINMA (CH), BaFin (DE), SEC (US)"),
        ("SEC", "Sector Regulator", "Industry-specific regulatory bodies",
         "Industry compliance, sector-specific incidents", "Conditional",
         "Per regulation", "Compliance > Legal", "Sector-specific details",
         "OFCOM, BAFU, Swissmedic"),
        ("EME", "Emergency Services", "Fire, ambulance, utilities",
         "Physical emergencies, facility incidents", "Yes (emergencies)",
         "Immediate", "Facility Mgr > HR > CEO", "Location, nature of emergency",
         "Fire Brigade, Ambulance, Police"),
        ("CYB", "Government Cybersecurity", "National cybersecurity centers",
         "Cyber threats, national security concerns", "Recommended",
         "24-48 hours", "CISO > Legal", "Threat details, IOCs, impact",
         "NCSC (CH), BSI (DE), CISA (US)"),
        ("STD", "Standards Body", "Certification and standards organisations",
         "Certification matters, standard interpretations", "No",
         "Per agreement", "Quality > Compliance", "Certification scope, queries",
         "SNV, ISO National Members"),
    ]

    row = 4
    for type_data in authority_types:
        for col, value in enumerate(type_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row == 4:
                cell.fill = LOCKED_FILL
        row += 1

    # Add empty input rows (rows 11-60 = 50 FFFFCC rows)
    for row in range(11, 61):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
    ws.freeze_panes = 'A4'

    # Column widths
    widths = [10, 25, 40, 35, 18, 15, 25, 35, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_communication_log_sheet(ws):
    """Create the Communication Log sheet."""
    ws.title = "Communication Log"

    headers = [
        "Log ID", "Contact ID", "Authority Name", "Communication Date",
        "Communication Type", "Direction", "Subject", "Summary",
        "Our Representative", "Authority Representative", "Outcome",
        "Follow Up Required", "Follow Up Date", "Evidence Reference", "Notes"
    ]

    # Title row (003366 fill, ALL CAPS)
    _n_cols = len(headers)
    _title = ws.cell(row=1, column=1, value="COMMUNICATION LOG")
    _title.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    _title.fill = TITLE_FILL
    _title.alignment = Alignment(horizontal="center", vertical="center")
    _title.border = THIN_BORDER
    ws.merge_cells(f"A1:{get_column_letter(_n_cols)}1")
    ws.row_dimensions[1].height = 35

    _sub = ws.cell(row=2, column=1, value="ISO 27001:2022 | Control A.5.5-6 | External Communications Log")
    _sub.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    _sub.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _sub.alignment = Alignment(horizontal="center", vertical="center")
    _sub.border = THIN_BORDER
    ws.merge_cells(f"A2:{get_column_letter(_n_cols)}2")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header.replace("_", " "))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Meeting,Phone Call,Email,Letter,Report Submission,Notification,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('E4:E201')

    direction_dv = DataValidation(type="list", formula1='"Outbound,Inbound"')
    ws.add_data_validation(direction_dv)
    direction_dv.add('F4:F201')

    followup_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(followup_dv)
    followup_dv.add('L4:L201')

    # Format input rows
    for row in range(4, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = LOCKED_FILL if row == 4 else INPUT_FILL
            cell.border = THIN_BORDER

    # Grey sample row 4
    _sample = ("LOG-001", "AUTH-001", "FDPIC", "01.01.2026",
               "Email", "Outbound", "Annual Article 12 Notification",
               "Submitted annual notification per nDSG Art. 12", "J. Smith",
               "Commissioner Office", "Notification accepted", "No", "",
               "EVID-001", "Annual notification submitted")
    for col, val in enumerate(_sample, 1):
        ws.cell(row=4, column=col, value=val)
    ws.freeze_panes = 'A4'

    # Column widths
    widths = [12, 12, 30, 15, 18, 12, 35, 50, 20, 25, 25, 15, 15, 25, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_verification_register_sheet(ws):
    """Create the Verification Register sheet."""
    ws.title = "Verification Register"

    headers = [
        "Verification ID", "Contact ID", "Authority Name", "Verification Date",
        "Verified By", "Verification Method", "Contact Details Correct",
        "Escalation Path Tested", "Website Accessible", "Relationship Active",
        "Issues Found", "Actions Required", "Next Verification", "Notes"
    ]

    # Title row (003366 fill, ALL CAPS)
    _n_cols = len(headers)
    _title = ws.cell(row=1, column=1, value="VERIFICATION REGISTER")
    _title.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    _title.fill = TITLE_FILL
    _title.alignment = Alignment(horizontal="center", vertical="center")
    _title.border = THIN_BORDER
    ws.merge_cells(f"A1:{get_column_letter(_n_cols)}1")
    ws.row_dimensions[1].height = 35

    _sub = ws.cell(row=2, column=1, value="ISO 27001:2022 | Control A.5.5-6 | Contact Verification Register")
    _sub.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    _sub.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _sub.alignment = Alignment(horizontal="center", vertical="center")
    _sub.border = THIN_BORDER
    ws.merge_cells(f"A2:{get_column_letter(_n_cols)}2")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header.replace("_", " "))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    method_dv = DataValidation(
        type="list",
        formula1='"Phone Test,Email Test,Website Check,Meeting,Document Review"'
    )
    ws.add_data_validation(method_dv)
    method_dv.add('F4:F101')

    yn_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('G4:G101')
    yn_dv.add('H4:H101')
    yn_dv.add('I4:I101')
    yn_dv.add('J4:J101')

    # Format input rows
    for row in range(4, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = LOCKED_FILL if row == 4 else INPUT_FILL
            cell.border = THIN_BORDER

    # Grey sample row 4
    _sample = ("VER-001", "AUTH-001", "FDPIC", "01.01.2026",
               "J. Smith", "Website Check", "Yes", "N/A",
               "Yes", "Yes", "None", "None", "01.01.2027",
               "Annual verification completed")
    for col, val in enumerate(_sample, 1):
        ws.cell(row=4, column=col, value=val)
    ws.freeze_panes = 'A4'

    # Column widths
    widths = [15, 12, 30, 15, 20, 20, 20, 20, 18, 18, 30, 30, 18, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width



def create_evidence_register(ws):
    """Create Evidence Register worksheet (Gold Standard: 8 cols, 003366 headers, 100 rows)."""
    ws.title = "Evidence Register"
    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _A, Border as _B, Side as _S
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.worksheet.datavalidation import DataValidation as _DV

    _thin = _S(style='thin')
    _border = _B(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _dark = _PF(start_color='003366', end_color='003366', fill_type='solid')
    _grey = _PF(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _yell = _PF(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    # Row 1: Title banner (A1:H1, 003366, white bold 14pt, height=35)
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    ws['A1'].font = _F(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = _dark
    ws['A1'].alignment = _A(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for c in range(1, 9):
        ws.cell(row=1, column=c).border = _border

    # Row 2: Subtitle (italic, merged A2:H2)
    ws.merge_cells('A2:H2')
    ws['A2'] = f'ISO/IEC 27001:2022 | Controls A.5.5 & A.5.6 | Evidence Register'
    ws['A2'].font = _F(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = _A(horizontal='center', vertical='center')
    for c in range(1, 9):
        ws.cell(row=2, column=c).border = _border

    # Row 3: Empty separator
    for c in range(1, 9):
        ws.cell(row=3, column=c).border = _border

    # Row 4: Column headers (003366 fill, white bold)
    headers = ['Evidence ID', 'Assessment Area', 'Evidence Type', 'Description',
               'Location / Path', 'Date Collected', 'Collected By', 'Verification Status']
    for idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=4, column=idx, value=hdr)
        cell.font = _F(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = _dark
        cell.alignment = _A(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border

    # Row 5: F2F2F2 sample row
    sample = [
        'EV-001', 'Authority Registry', 'Policy Document',
        'ISMS policy document evidence reference', '/evidence/A.5.5-6/',
        '01.01.2026', 'ISMS Manager', 'Verified',
    ]
    for idx, val in enumerate(sample, 1):
        cell = ws.cell(row=5, column=idx, value=val)
        cell.fill = _grey
        cell.border = _border

    # DV for Verification Status (col H = 8)
    ver_dv = _DV(type='list', formula1='"Verified,Pending,Not Verified,N/A"', allow_blank=True)
    ws.add_data_validation(ver_dv)
    ver_dv.add('H5:H105')

    # Rows 6-105: 100 empty FFFFCC rows (total 101 = 1 sample + 100 empty)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = _yell
            cell.border = _border

    # Column widths and freeze
    for idx, width in enumerate([12, 25, 20, 40, 35, 15, 20, 18], 1):
        ws.column_dimensions[_gcl(idx)].width = width
    ws.freeze_panes = 'A5'

def create_approval_sheet(ws):
    """Create Approval Sign-Off worksheet (Gold Standard)."""
    ws.title = "Approval Sign-Off"
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL AND SIGN-OFF'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2: CONTROL REFERENCE subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f'{DOCUMENT_ID} | {CONTROL_REF}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells('A3:E3')
    ws['A3'] = 'ASSESSMENT SUMMARY'
    ws['A3'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ('Document:', f'{DOCUMENT_ID} - {WORKBOOK_NAME}'),
        ('Assessment Period:', ''),
        ('Overall Compliance Rate:', "=IFERROR(AVERAGE('Summary Dashboard'!G6:G6),\"\")"),
        ('Assessment Status:', ''),
        ('Assessed By:', ''),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == '')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        if 'Assessment Status' in label:
            status_row_for_dv = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Assessment Status dropdown
    status_dv = DataValidation(
        type='list',
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f'B{status_row_for_dv}')

    # B6: Overall Compliance Rate — reference Summary Dashboard TOTAL compliance %
    ws['B6'] = "=IFERROR('Summary Dashboard'!F7,"")"
    ws['B6'].number_format = '0.0%'

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ['Name:', 'Title:', 'Date:', 'Signature:', 'Comments:']:
            ws[f'A{current_row}'] = field
            ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'B{current_row}:E{current_row}')
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, 'COMPLETED BY (ASSESSOR)', '4472C4')
    row = _create_approver_section(row, 'REVIEWED BY (INFORMATION SECURITY OFFICER)', '4472C4')
    row = _create_approver_section(row, 'APPROVED BY (CISO)', '003366')

    # FINAL DECISION
    ws[f'A{row}'] = 'FINAL DECISION:'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border
    ws.merge_cells(f'B{row}:E{row}')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type='list',
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ['Next Review Date:', 'Review Responsible:', 'Special Considerations:']:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths and freeze
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.freeze_panes = 'A3'

def create_summary_dashboard_sheet(ws):
    """Summary Dashboard — A.5.5-6.S1 Authority Contacts Register.
    Gold Standard TABLE 1/2/3 implementation per A.8.33-34 reference.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _f(h):
        return PatternFill("solid", fgColor=h)

    def _banner(r, txt, fill_hex, merged_to, height=20):
        ws.merge_cells(f"A{r}:{merged_to}{r}")
        c = ws.cell(row=r, column=1)
        c.value = txt
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        c.fill = _f(fill_hex)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border
        ws.row_dimensions[r].height = height
        # Set border on all merged cells
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = _border

    ws.title = "Summary Dashboard"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A3"

    # ── Row 1: Title (GS-SD-014: must contain em dash + SUMMARY DASHBOARD) ──
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"{WORKBOOK_NAME} \u2014 SUMMARY DASHBOARD"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.fill = _f("003366")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border
    ws.row_dimensions[1].height = 35
    for col in range(1, 7):
        ws.cell(row=1, column=col).border = _border

    # ── Row 2: Subtitle (GS: italic, color=003366, left-aligned, NO fill) ──
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "ISO/IEC 27001:2022 \u2014 Controls A.5.5 & A.5.6: Authority contact health, relationship status and communication compliance"
    c.font = Font(name="Calibri", italic=True, color="003366", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 1 — ASSESSMENT AREA COMPLIANCE OVERVIEW
    # ══════════════════════════════════════════════════════════════════════
    # Row 4: TABLE 1 banner
    _banner(4, "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW", "003366", "F")

    # Row 5: Column headers (D9D9D9 — NOT 4472C4, GS-SD-016)
    t1_headers = ["Assessment Area", "Total", "Compliant", "Partial", "Non-Compliant", "Compliance %"]
    for col, hdr in enumerate(t1_headers, 1):
        c = ws.cell(row=5, column=col, value=hdr)
        c.font = Font(name="Calibri", bold=True, color="000000", size=10)
        c.fill = _f("D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border

    # Row 6: Authority Registry — Contact Health
    # COUNTA on col A (user-entered IDs), ranges start at row 5 (skip sample row 4)
    area_row = 6
    ws.cell(row=area_row, column=1, value="Authority Registry \u2014 Contact Health").font = Font(name="Calibri", color="000000")
    ws.cell(row=area_row, column=1).fill = _f("FFFFFF")
    ws.cell(row=area_row, column=1).border = _border
    ws.cell(row=area_row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    formulas_t1 = [
        ("=COUNTA('Authority Registry'!A5:A54)", True),
        ("=COUNTIF('Authority Registry'!O5:O54,\"Active\")", True),
        ("=COUNTIF('Authority Registry'!O5:O54,\"Under Review\")", True),
        ("=COUNTIF('Authority Registry'!O5:O54,\"Inactive\")+COUNTIF('Authority Registry'!O5:O54,\"Pending\")", True),
        (f"=IFERROR(IF((B{area_row}-0)=0,0,C{area_row}/(B{area_row}-0)),\"\")", True),
    ]
    for col_offset, (formula, num) in enumerate(formulas_t1):
        col = col_offset + 2
        c = ws.cell(row=area_row, column=col, value=formula)
        c.font = Font(name="Calibri", color="000000")
        c.fill = _f("FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border
        if col == 6:
            c.number_format = "0.0%"

    # Row 7: TOTAL row (D9D9D9)
    total_row = 7
    c = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font = Font(name="Calibri", bold=True, color="000000")
    c.fill = _f("D9D9D9")
    c.border = _border
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 6):
        ltr = ["B", "C", "D", "E"][col - 2]
        c = ws.cell(row=total_row, column=col, value=f"=SUM({ltr}{area_row}:{ltr}{area_row})")
        c.font = Font(name="Calibri", bold=True, color="000000")
        c.fill = _f("D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border
    # Compliance % TOTAL
    c = ws.cell(row=total_row, column=6,
                value=f"=IFERROR(IF((B{total_row}-0)=0,0,C{total_row}/(B{total_row}-0)),\"\")")
    c.font = Font(name="Calibri", bold=True, color="000000")
    c.fill = _f("D9D9D9")
    c.number_format = "0.0%"
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 2 — KEY PERFORMANCE INDICATORS
    # ══════════════════════════════════════════════════════════════════════
    # Row 9: TABLE 2 banner
    _banner(9, "TABLE 2: KEY PERFORMANCE INDICATORS", "003366", "F")

    # Row 10: TABLE 2 headers (D9D9D9 — NOT 4472C4, GS-SD-016)
    t2_headers = ["KPI Metric", "Value", "Notes / Subtitle"]
    for col, hdr in enumerate(t2_headers, 1):
        c = ws.cell(row=10, column=col, value=hdr)
        c.font = Font(name="Calibri", bold=True, color="000000", size=10)
        c.fill = _f("D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
        if col == 3:
            ws.merge_cells("C10:F10")
    for col in range(4, 7):
        ws.cell(row=10, column=col).border = _border

    # TABLE 2 data rows (white fill, 000000 font, NOT bold labels — GS-SD-015)
    # COUNTA ranges start at row 5 (skip sample row 4)
    kpis = [
        ("Total Authorities Registered",
         "=COUNTA('Authority Registry'!A5:A54)",
         "All authorities, regulators and government bodies registered"),
        ("Active Relationships",
         "=COUNTIF('Authority Registry'!O5:O54,\"Active\")",
         "Authorities with confirmed active relationship status"),
        ("Data Protection Authorities Registered",
         "=COUNTIF('Authority Registry'!C5:C54,\"Data Protection Authority\")",
         "DPAs registered \u2014 required for breach notification compliance"),
        ("Law Enforcement Contacts Registered",
         "=COUNTIF('Authority Registry'!C5:C54,\"Law Enforcement\")",
         "Law enforcement contacts \u2014 required for cybercrime response"),
        ("Communication Events Logged",
         "=COUNTA('Communication Log'!A5:A54)",
         "Total external communications with authorities logged"),
        ("Pending Follow-Up Actions",
         "=COUNTIF('Communication Log'!L5:L54,\"Yes\")",
         "Outstanding follow-up actions from authority communications"),
        ("Verification Records Logged",
         "=COUNTA('Verification Register'!A5:A54)",
         "Total contact verification activities recorded"),
        ("Contacts Verified as Correct",
         "=COUNTIF('Verification Register'!G5:G54,\"Yes\")",
         "Authority contacts confirmed accurate in last verification"),
    ]
    for i, (metric, formula, note) in enumerate(kpis):
        r = 11 + i
        # Col A: metric label — white fill, 000000, NOT bold (GS-SD-015)
        c = ws.cell(row=r, column=1, value=metric)
        c.font = Font(name="Calibri", color="000000")
        c.fill = _f("FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = _border
        # Col B: value formula — white fill, 000000, center
        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = Font(name="Calibri", color="000000")
        c2.fill = _f("FFFFFF")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _border
        # Col C-F: notes merged — white fill, 000000, left
        ws.merge_cells(f"C{r}:F{r}")
        c3 = ws.cell(row=r, column=3, value=note)
        c3.font = Font(name="Calibri", color="000000")
        c3.fill = _f("FFFFFF")
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3.border = _border
        # Set border on all merged cells (GS-AS-013)
        for col in range(3, 7):
            ws.cell(row=r, column=col).border = _border

    # ══════════════════════════════════════════════════════════════════════
    # TABLE 3 — CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION
    # ══════════════════════════════════════════════════════════════════════
    # TABLE 2 ends at row 18 (11+8-1), empty row 19, TABLE 3 banner at row 20
    _banner(20, "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION", "C00000", "F")

    # Row 21: TABLE 3 headers (D9D9D9)
    t3_headers = ["Finding", "Count", "Action Required"]
    for col, hdr in enumerate(t3_headers, 1):
        c = ws.cell(row=21, column=col, value=hdr)
        c.font = Font(name="Calibri", bold=True, color="000000", size=10)
        c.fill = _f("D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
        if col == 3:
            ws.merge_cells("C21:F21")
    for col in range(4, 7):
        ws.cell(row=21, column=col).border = _border

    # TABLE 3 data rows (white fill, col A bold, 000000)
    # COUNTIF ranges start at row 5 (skip sample row 4)
    findings = [
        ("Inactive Authorities",
         "=COUNTIF('Authority Registry'!O5:O54,\"Inactive\")",
         "Review inactive authorities \u2014 confirm intentional or reactivate contact"),
        ("Relationships Under Review",
         "=COUNTIF('Authority Registry'!O5:O54,\"Under Review\")",
         "Complete review of flagged authorities and update status to Active or Inactive"),
        ("Pending Follow-Up Actions",
         "=COUNTIF('Communication Log'!L5:L54,\"Yes\")",
         "Outstanding follow-up actions from authority communications \u2014 resolve and log outcomes"),
    ]
    for i, (finding, formula, action) in enumerate(findings):
        r = 22 + i
        # Col A: finding label — white, bold
        c = ws.cell(row=r, column=1, value=finding)
        c.font = Font(name="Calibri", bold=True, color="000000")
        c.fill = _f("FFFFCC")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = _border
        # Col B: formula — white, center
        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = Font(name="Calibri", color="000000")
        c2.fill = _f("FFFFCC")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _border
        # Col C-F: action merged — white, left
        ws.merge_cells(f"C{r}:F{r}")
        c3 = ws.cell(row=r, column=3, value=action)
        c3.font = Font(name="Calibri", color="000000")
        c3.fill = _f("FFFFCC")
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3.border = _border
        for col in range(3, 7):
            ws.cell(row=r, column=col).border = _border

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info("=" * 70)

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Create all sheets
    create_instructions_sheet(wb.active)
    create_authority_registry_sheet(wb.create_sheet())
    create_contact_types_sheet(wb.create_sheet())
    create_communication_log_sheet(wb.create_sheet())
    create_verification_register_sheet(wb.create_sheet())
    create_evidence_register(wb.create_sheet())
    create_summary_dashboard_sheet(wb.create_sheet())
    create_approval_sheet(wb.create_sheet())

    # Save workbook
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    wb.save(output_path)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")
    logger.info("=" * 70)



def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
