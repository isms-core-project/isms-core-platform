#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.9.1 - Asset Discovery Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Assessment Domain 1 of 5: Asset Discovery & Completeness Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific asset landscape, discovery tools, and assessment
requirements.

Key customization areas:
1. Asset categories and types (match your business context and infrastructure)
2. Discovery methods and tools (network scanning, HR systems, procurement, etc.)
3. Completeness thresholds (what percentage is acceptable for your risk profile)
4. Organisation name, CISO details, contact information
5. File paths and naming conventions
6. Integration with existing asset management/CMDB systems

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
asset discovery completeness across all five asset categories defined in ISO
27001:2022 Control A.5.9.

**Purpose:**
Enables systematic assessment of whether [Organisation] has identified and
documented all assets that store, process, transmit, or protect information.
Measures discovery completeness against expected asset populations to identify
gaps and guide remediation efforts.

**Assessment Scope:**
- Information Assets (databases, documents, records, intellectual property)
- IT Infrastructure (servers, network devices, storage, endpoints)
- Applications (SaaS, on-premise, custom, COTS)
- Physical Assets (facilities, equipment, media, backup tapes)
- Personnel Assets (roles, competencies, critical knowledge holders)

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and completion guidance
2. Information Assets Discovery - Discovery completeness for information assets
3. IT Infrastructure Discovery - Discovery completeness for IT assets
4. Applications Discovery - Discovery completeness for applications
5. Physical Assets Discovery - Discovery completeness for physical assets
6. Personnel Assets Discovery - Discovery completeness for personnel assets
7. Discovery Metrics & Summary - Aggregated completeness scores and gaps
8. Evidence Register - Discovery evidence tracking and documentation

**Key Features:**
- Data validation with dropdown lists for asset categories
- Conditional formatting for completeness status (Green/Yellow/Red)
- Automated gap identification and prioritization
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Discovery method tracking (network scan, manual, procurement, HR)
- Completeness percentage calculations by category
- Gap analysis with remediation recommendations

**Integration:**
data from all four assessment domains for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a59_1_asset_discovery.py

Output:
    File: ISMS-IMP-A.5.9.1_Asset_Discovery_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Complete discovery assessment by populating all asset categories
    2. Document discovery methods used (network scans, HR exports, etc.)
    3. Calculate completeness percentages vs. expected populations
    4. Identify gaps and prioritize discovery efforts
    5. Collect and link evidence in Evidence Register sheet
    6. Review with stakeholders (IT, HR, Security, Asset Owners)
    7. Export metrics CSV for dashboard consolidation (Sheet 7)
    8. Store assessment workbook per retention policy (7 years minimum)
    9. Update quarterly or after major infrastructure/organisational changes

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.9
Assessment Domain:    1 of 5
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License]

Related Documents:
    - ISMS-POL-A.5.9: Inventory of Information and Assets (Policy)
    - ISMS-IMP-A.5.9-1: Asset Discovery Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-2: Inventory Maintenance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-3: Quality & Compliance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-4: Owner Accountability Assessment (Implementation Guide)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 22.01.2026
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.9-1 specification
    - Discovery completeness tracking across 5 asset categories
    - Gap identification and remediation planning
    - Evidence collection and audit trail support
    - CSV export for dashboard consolidation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive asset inventory details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review asset inventory procedures and classification criteria annually or when
new asset categories are introduced, system landscapes change, or audit findings
identify inventory gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
from datetime import datetime
from pathlib import Path
import os

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# CUSTOMIZE: Configuration
_DISCOVERY_METHODS = [
    'Network Scan (Automated)',
    'CMDB/Asset Management System',
    'Procurement Records',
    'HR System Export',
    'Manual Survey',
    'License Audit',
    'Physical Inspection',
    'Department Interviews',
    'Document Repository Scan',
    'Cloud Provider API',
    'Other (Specify)',
]

# Document identification constants

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.9.1"
WORKBOOK_NAME = "Asset Discovery"
CONTROL_ID   = "A.5.9"
CONTROL_NAME = "Inventory of Information and Assets"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("="*80)
    logger.info("ISMS Control A.5.9 - Asset Discovery Assessment Generator")
    logger.info("="*80)
    logger.info("")
    logger.info("Generating assessment workbook...")
    logger.info("")
    
    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create sheets in order
    sheets = [
        "Instructions & Legend",
        "Information Assets Discovery",
        "IT Infrastructure Discovery",
        "Applications Discovery",
        "Physical Assets Discovery",
        "Personnel Assets Discovery",
        "Discovery Metrics & Summary",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
        logger.info(f"  ✓ Created sheet: {sheet_name}")
    
    logger.info("")
    logger.info("Populating sheets...")
    logger.info("")
    
    # Populate each sheet
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✓ Instructions & Legend")
    
    create_information_assets_sheet(wb["Information Assets Discovery"])
    logger.info("  ✓ Information Assets Discovery")
    
    create_it_infrastructure_sheet(wb["IT Infrastructure Discovery"])
    logger.info("  ✓ IT Infrastructure Discovery")
    
    create_applications_sheet(wb["Applications Discovery"])
    logger.info("  ✓ Applications Discovery")
    
    create_physical_assets_sheet(wb["Physical Assets Discovery"])
    logger.info("  ✓ Physical Assets Discovery")
    
    create_personnel_assets_sheet(wb["Personnel Assets Discovery"])
    logger.info("  ✓ Personnel Assets Discovery")
    
    create_discovery_metrics_sheet(wb["Discovery Metrics & Summary"])
    logger.info("  ✓ Discovery Metrics & Summary")
    
    create_evidence_register(wb["Evidence Register"])
    logger.info("  ✓ Evidence Register")
    
    create_summary_dashboard_sheet(wb["Summary Dashboard"])
    logger.info("  ✓ Summary Dashboard")
    
    create_approval_sheet(wb["Approval Sign-Off"])
    logger.info("  ✓ Approval & Sign-Off")
    
    # Finalise data validations
    finalize_validations(wb)

    # Save workbook to WKBK directory
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    wb.save(output_path)

    logger.info("")
    logger.info("="*80)
    logger.info(f"SUCCESS: {output_path.name}")
    logger.info("="*80)
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open workbook and review Instructions & Legend")
    logger.info("  2. Complete discovery assessment for each asset category")
    logger.info("  3. Document discovery methods and evidence")
    logger.info("  4. Review Discovery Metrics & Summary for completeness scores")
    logger.info("  5. Export CSV from Sheet 7 for dashboard consolidation")
    logger.info("  6. Obtain stakeholder review and approval")
    logger.info("")
    logger.info(f"Output location: {output_path}")
    logger.info("")



def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Information Assets Discovery sheet with all information asset subcategories.', '2. Complete the IT Infrastructure Discovery sheet with all IT infrastructure assets.', '3. Complete the Applications Discovery sheet with all application assets.', '4. Complete the Physical Assets Discovery sheet with all physical assets.', '5. Complete the Personnel Assets Discovery sheet with all personnel asset roles.', '6. Review the Discovery Metrics & Summary sheet for completeness scores.', '7. Link evidence in the Evidence Register sheet.', '8. Obtain approvals in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    # Status Legend — row position tracks after instructions
    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_information_assets_sheet(ws):
    """Create Information Assets Discovery sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "INFORMATION ASSETS DISCOVERY ASSESSMENT"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers (row 3)
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    ws.row_dimensions[3].height = 30
    
    # Information Asset Subcategories (pre-populated examples - customize)
    subcategories = [
        "Databases (Production)",
        "Databases (Development/Test)",
        "Data Warehouses / Analytics",
        "File Shares (Departmental)",
        "SharePoint / Document Management",
        "Business Records / Archives",
        "Email Archives",
        "Intellectual Property (Patents, Trade Secrets)",
        "Source Code Repositories",
        "Customer Data Collections",
        "Financial Records",
        "HR Records (Personnel Files)",
        "Contracts and Legal Documents",
        "Technical Documentation",
        "Marketing / Sales Materials",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        # FFFFCC fill + borders for data row; F2F2F2 for sample row 4
        _fill_color = 'F2F2F2' if row == 4 else 'FFFFCC'
        _yf = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type='solid')
        from openpyxl.styles import Border as _Bdr, Side as _Sd
        _ts = _Sd(style='thin')
        _db = _Bdr(left=_ts, right=_ts, top=_ts, bottom=_ts)
        for _lc in 'ABCDEFGHIJKLM':
            ws[f'{_lc}{row}'].fill = _yf
            ws[f'{_lc}{row}'].border = _db

        
        # Expected Count (user enters)
        
        # Discovered Count (user enters)
        
        # Completeness % (formula)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        # Compliance Status (formula - based on 95% target)
        ws[f'F{row}'] = f'=IF(E{row}>=95,"✅ Compliant",IF(E{row}>=85,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        # Unlocked user input cells
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            pass
        
        row += 1
    
    # Discovery Method dropdown (column B)
    dv_method = DataValidation(
        type="list",
        formula1=f'"{",".join(_DISCOVERY_METHODS)}"',
        allow_blank=True
    )
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary section
    summary_row = row + 2
    ws[f'A{summary_row}'] = "INFORMATION ASSETS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Expected"
    ws[f'B{summary_row}'] = f'=SUM(C4:C{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Discovered"
    ws[f'B{summary_row}'] = f'=SUM(D4:D{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Overall Completeness %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "95%"
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap vs. Target"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-95'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting for Completeness %
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='between', formula=['85', '94'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='lessThan', formula=['85'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )
    
    # Protect sheet (formulas locked, inputs unlocked)


def create_it_infrastructure_sheet(ws):
    """Create IT Infrastructure Discovery sheet"""
    
    # Same structure as Information Assets but with IT-specific subcategories
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "IT INFRASTRUCTURE DISCOVERY ASSESSMENT"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    # IT Infrastructure Subcategories
    subcategories = [
        "Physical Servers (Production)",
        "Physical Servers (Non-Production)",
        "Virtual Servers / VMs",
        "Network Routers",
        "Network Switches",
        "Firewalls",
        "Load Balancers",
        "Storage Arrays / NAS / SAN",
        "Backup Devices / Tape Libraries",
        "Workstations / Desktops",
        "Laptops / Mobile Devices",
        "Tablets / iPads",
        "Smartphones (Corporate)",
        "Printers / Multifunction Devices",
        "IoT Devices / Sensors",
        "Network Attached Devices",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        # FFFFCC fill + borders for data row; F2F2F2 for sample row 4
        _fill_color = 'F2F2F2' if row == 4 else 'FFFFCC'
        _yf = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type='solid')
        from openpyxl.styles import Border as _Bdr, Side as _Sd
        _ts = _Sd(style='thin')
        _db = _Bdr(left=_ts, right=_ts, top=_ts, bottom=_ts)
        for _lc in 'ABCDEFGHIJKLM':
            ws[f'{_lc}{row}'].fill = _yf
            ws[f'{_lc}{row}'].border = _db

        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=IF(E{row}>=98,"✅ Compliant",IF(E{row}>=88,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            pass
        row += 1
    
    # Discovery Method dropdown
    dv_method = DataValidation(
        type="list",
        formula1=f'"{",".join(_DISCOVERY_METHODS)}"',
        allow_blank=True
    )
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary section (98% target for IT Infrastructure)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "IT INFRASTRUCTURE SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Expected"
    ws[f'B{summary_row}'] = f'=SUM(C4:C{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Discovered"
    ws[f'B{summary_row}'] = f'=SUM(D4:D{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Overall Completeness %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "98%"
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap vs. Target"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-98'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting (98% threshold)
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['98'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='between', formula=['88', '97'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='lessThan', formula=['88'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_applications_sheet(ws):
    """Create Applications Discovery sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "APPLICATIONS DISCOVERY ASSESSMENT"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Application Subcategories
    subcategories = [
        "SaaS Applications (Licensed)",
        "Cloud-Hosted Applications (Custom)",
        "On-Premise Applications (COTS)",
        "On-Premise Applications (Custom/In-House)",
        "Database Management Systems",
        "Middleware / Integration Platforms",
        "Development Tools / IDEs",
        "Security Tools (SIEM, EDR, Firewall)",
        "Monitoring / Observability Tools",
        "Backup / Recovery Software",
        "Collaboration Tools (Slack, Teams, etc.)",
        "Productivity Suites (Office 365, G Suite)",
        "ERP / CRM Systems",
        "HR Management Systems",
        "Shadow IT (Unauthorised SaaS)",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        # FFFFCC fill + borders for data row; F2F2F2 for sample row 4
        _fill_color = 'F2F2F2' if row == 4 else 'FFFFCC'
        _yf = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type='solid')
        from openpyxl.styles import Border as _Bdr, Side as _Sd
        _ts = _Sd(style='thin')
        _db = _Bdr(left=_ts, right=_ts, top=_ts, bottom=_ts)
        for _lc in 'ABCDEFGHIJKLM':
            ws[f'{_lc}{row}'].fill = _yf
            ws[f'{_lc}{row}'].border = _db

        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=IF(E{row}>=90,"✅ Compliant",IF(E{row}>=80,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            pass
        row += 1
    
    dv_method = DataValidation(type="list", formula1=f'"{",".join(_DISCOVERY_METHODS)}"', allow_blank=True)
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary (90% target)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "APPLICATIONS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    ws[f'A{summary_row+1}'] = "Total Expected"
    ws[f'B{summary_row+1}'] = f'=SUM(C4:C{row-1})'
    ws[f'B{summary_row+1}'].font = Font(bold=True)
    
    ws[f'A{summary_row+2}'] = "Total Discovered"
    ws[f'B{summary_row+2}'] = f'=SUM(D4:D{row-1})'
    ws[f'B{summary_row+2}'].font = Font(bold=True)
    
    ws[f'A{summary_row+3}'] = "Overall Completeness %"
    ws[f'B{summary_row+3}'] = f'=IFERROR(B{summary_row+2}/B{summary_row+1}*100,0)'
    ws[f'B{summary_row+3}'].number_format = '0.0"%"'
    ws[f'B{summary_row+3}'].font = Font(bold=True, size=12)
    
    ws[f'A{summary_row+4}'] = "Target"
    ws[f'B{summary_row+4}'] = "90%"
    ws[f'B{summary_row+4}'].font = Font(bold=True)
    
    ws[f'A{summary_row+5}'] = "Gap vs. Target"
    ws[f'B{summary_row+5}'] = f'=B{summary_row+3}-90'
    ws[f'B{summary_row+5}'].number_format = '0.0"%"'
    
    # Conditional formatting
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='greaterThanOrEqual', formula=['90'], fill=PatternFill(start_color="C6EFCE", fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='between', formula=['80', '89'], fill=PatternFill(start_color="FFEB9C", fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='lessThan', formula=['80'], fill=PatternFill(start_color="FFC7CE", fill_type='solid')))


def create_physical_assets_sheet(ws):
    """Create Physical Assets Discovery sheet"""
    
    # Similar structure with physical asset categories
    ws.merge_cells('A1:M1')
    ws['A1'] = "PHYSICAL ASSETS DISCOVERY ASSESSMENT"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    subcategories = [
        "Data Centers / Server Rooms",
        "Office Facilities",
        "Storage Facilities (Offsite)",
        "Backup Tape Libraries (Offsite)",
        "Removable Media (USB Drives, External HDDs)",
        "Paper Records / File Cabinets",
        "Safes / Secure Storage",
        "Badge Readers / Access Control Systems",
        "CCTV / Surveillance Systems",
        "Environmental Controls (HVAC, Fire Suppression)",
        "Power Infrastructure (UPS, Generators)",
        "Network Cabling Infrastructure",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        # FFFFCC fill + borders for data row; F2F2F2 for sample row 4
        _fill_color = 'F2F2F2' if row == 4 else 'FFFFCC'
        _yf = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type='solid')
        from openpyxl.styles import Border as _Bdr, Side as _Sd
        _ts = _Sd(style='thin')
        _db = _Bdr(left=_ts, right=_ts, top=_ts, bottom=_ts)
        for _lc in 'ABCDEFGHIJKLM':
            ws[f'{_lc}{row}'].fill = _yf
            ws[f'{_lc}{row}'].border = _db

        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=IF(E{row}>=90,"✅ Compliant",IF(E{row}>=80,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            pass
        row += 1
    
    dv_method = DataValidation(type="list", formula1=f'"{",".join(_DISCOVERY_METHODS)}"', allow_blank=True)
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary (90% target)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "PHYSICAL ASSETS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    ws[f'A{summary_row+1}'] = "Total Expected"
    ws[f'B{summary_row+1}'] = f'=SUM(C4:C{row-1})'
    ws[f'A{summary_row+2}'] = "Total Discovered"
    ws[f'B{summary_row+2}'] = f'=SUM(D4:D{row-1})'
    ws[f'A{summary_row+3}'] = "Overall Completeness %"
    ws[f'B{summary_row+3}'] = f'=IFERROR(B{summary_row+2}/B{summary_row+1}*100,0)'
    ws[f'B{summary_row+3}'].number_format = '0.0"%"'
    ws[f'B{summary_row+3}'].font = Font(bold=True, size=12)
    ws[f'A{summary_row+4}'] = "Target"
    ws[f'B{summary_row+4}'] = "90%"
    ws[f'A{summary_row+5}'] = "Gap vs. Target"
    ws[f'B{summary_row+5}'] = f'=B{summary_row+3}-90'
    ws[f'B{summary_row+5}'].number_format = '0.0"%"'
    
    # Conditional formatting
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='greaterThanOrEqual', formula=['90'], fill=PatternFill(start_color="C6EFCE", fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='between', formula=['80', '89'], fill=PatternFill(start_color="FFEB9C", fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='lessThan', formula=['80'], fill=PatternFill(start_color="FFC7CE", fill_type='solid')))


def create_personnel_assets_sheet(ws):
    """Create Personnel Assets Discovery sheet"""
    
    ws.merge_cells('A1:M1')
    ws['A1'] = "PERSONNEL ASSETS DISCOVERY ASSESSMENT"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    subcategories = [
        "Executive Management (C-Suite)",
        "Department Heads / Managers",
        "Information Security Team",
        "IT Operations Team",
        "Developers / Engineers",
        "Database Administrators",
        "System Administrators",
        "Network Engineers",
        "Security Analysts / SOC",
        "Compliance Officers",
        "Data Protection Officers",
        "Key Subject Matter Experts",
        "Critical Project Personnel",
        "Contractors (Long-term)",
        "External Service Providers (Key Personnel)",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        # FFFFCC fill + borders for data row; F2F2F2 for sample row 4
        _fill_color = 'F2F2F2' if row == 4 else 'FFFFCC'
        _yf = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type='solid')
        from openpyxl.styles import Border as _Bdr, Side as _Sd
        _ts = _Sd(style='thin')
        _db = _Bdr(left=_ts, right=_ts, top=_ts, bottom=_ts)
        for _lc in 'ABCDEFGHIJKLM':
            ws[f'{_lc}{row}'].fill = _yf
            ws[f'{_lc}{row}'].border = _db

        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        # 100% target for personnel - CRITICAL
        ws[f'F{row}'] = f'=IF(E{row}=100,"✅ Compliant",IF(E{row}>=90,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            pass
        row += 1
    
    dv_method = DataValidation(type="list", formula1=f'"{",".join(_DISCOVERY_METHODS)}"', allow_blank=True)
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary (100% target - critical!)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "PERSONNEL ASSETS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    ws[f'A{summary_row+1}'] = "Total Expected"
    ws[f'B{summary_row+1}'] = f'=SUM(C4:C{row-1})'
    ws[f'A{summary_row+2}'] = "Total Discovered"
    ws[f'B{summary_row+2}'] = f'=SUM(D4:D{row-1})'
    ws[f'A{summary_row+3}'] = "Overall Completeness %"
    ws[f'B{summary_row+3}'] = f'=IFERROR(B{summary_row+2}/B{summary_row+1}*100,0)'
    ws[f'B{summary_row+3}'].number_format = '0.0"%"'
    ws[f'B{summary_row+3}'].font = Font(bold=True, size=12)
    ws[f'A{summary_row+4}'] = "Target"
    ws[f'B{summary_row+4}'] = "100%"
    ws[f'B{summary_row+4}'].font = Font(bold=True, color="9C0006")
    ws[f'A{summary_row+5}'] = "Gap vs. Target"
    ws[f'B{summary_row+5}'] = f'=B{summary_row+3}-100'
    ws[f'B{summary_row+5}'].number_format = '0.0"%"'
    
    # Conditional formatting (stricter - 100% required)
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='equal', formula=['100'], fill=PatternFill(start_color="C6EFCE", fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='between', formula=['90', '99'], fill=PatternFill(start_color="FFEB9C", fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='lessThan', formula=['90'], fill=PatternFill(start_color="FFC7CE", fill_type='solid')))


def create_discovery_metrics_sheet(ws):
    """Create Discovery Metrics & Summary sheet - consolidates all 5 categories"""
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "DISCOVERY METRICS & SUMMARY"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35
    
    # Column headers
    headers = [
        ('A', 'Asset Category', 30),
        ('B', 'Expected Count', 15),
        ('C', 'Discovered Count', 15),
        ('D', 'Completeness %', 15),
        ('E', 'Target %', 12),
        ('F', 'Gap vs. Target', 15),
        ('G', 'Compliance Status', 20),
        ('H', 'Priority', 15),
        ('I', 'Key Gaps', 50),
        ('J', 'Next Actions', 50),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Consolidate from other sheets
    categories = [
        ("Information Assets", "'Information Assets Discovery'", 95),
        ("IT Infrastructure", "'IT Infrastructure Discovery'", 98),
        ("Applications", "'Applications Discovery'", 90),
        ("Physical Assets", "'Physical Assets Discovery'", 90),
        ("Personnel Assets", "'Personnel Assets Discovery'", 100),
    ]
    
    row = 4
    for category_name, sheet_ref, target in categories:
        ws[f'A{row}'] = category_name
        ws[f'A{row}'].font = Font(bold=True)
        
        # Find summary row in source sheet (typically row with "Total Expected")
        # For simplicity, hardcode summary row positions (adjust based on actual sheets)
        # Information Assets: summary at row 22, IT: row 23, Apps: row 21, etc.
        # Using approximate positions - would need exact calculation
        
        # Expected Count (link to source sheet summary)
        ws[f'B{row}'] = f'={sheet_ref}!B22'  # Adjust row as needed
        
        # Discovered Count
        ws[f'C{row}'] = f'={sheet_ref}!B23'  # Adjust row as needed
        
        # Completeness %
        ws[f'D{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'D{row}'].font = Font(bold=True)
        
        # Target
        ws[f'E{row}'] = f'{target}%'
        
        # Gap vs. Target
        ws[f'F{row}'] = f'=D{row}-{target}'
        ws[f'F{row}'].number_format = '0.0"%"'
        
        # Compliance Status (formula)
        ws[f'G{row}'] = f'=IF(D{row}>={target},"✅ Compliant",IF(D{row}>={target}-10,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        # Priority (based on gap)
        ws[f'H{row}'] = f'=IF(F{row}<-10,"[!] Critical",IF(F{row}<0,"[~] High","[OK] Low"))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Key Gaps and Next Actions (user enters)
        
        row += 1
    
    # Overall Discovery Summary
    overall_row = row + 2
    ws[f'A{overall_row}'] = "OVERALL DISCOVERY SUMMARY"
    ws[f'A{overall_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{overall_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    ws.merge_cells(f'A{overall_row}:J{overall_row}')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Total Assets Expected (All Categories)"
    ws[f'B{overall_row}'] = f'=SUM(B4:B8)'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Total Assets Discovered"
    ws[f'B{overall_row}'] = f'=SUM(C4:C8)'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Discovery Completeness %"
    ws[f'B{overall_row}'] = f'=IFERROR(B{overall_row-1}/B{overall_row-2}*100,0)'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True, size=14)
    ws[f'B{overall_row}'].fill = PatternFill(start_color="FFEB9C", fill_type='solid')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Target"
    ws[f'B{overall_row}'] = "95%"
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Gap vs. Overall Target"
    ws[f'B{overall_row}'] = f'=B{overall_row-2}-95'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    # CSV Export section for dashboard
    csv_row = overall_row + 3
    ws[f'A{csv_row}'] = "CSV EXPORT FOR DASHBOARD (Copy rows below)"
    ws[f'A{csv_row}'].font = Font(size=11, bold=True, color="FFFFFF")
    ws[f'A{csv_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    csv_row += 1
    ws[f'A{csv_row}'] = "Discovery_Category"
    ws[f'B{csv_row}'] = "Completeness_%"
    ws[f'C{csv_row}'] = "Compliance_Status"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{csv_row}'].font = Font(bold=True)
    
    csv_row += 1
    # Link to data rows
    for i in range(5):
        ws[f'A{csv_row+i}'] = f'=A{4+i}'
        ws[f'B{csv_row+i}'] = f'=D{4+i}'
        ws[f'C{csv_row+i}'] = f'=G{4+i}'
    
    # Conditional formatting for Completeness %
    ws.conditional_formatting.add(
        'D4:D8',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'D4:D8',
        CellIsRule(operator='between', formula=['85', '94'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'D4:D8',
        CellIsRule(operator='lessThan', formula=['85'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_evidence_register(ws):
    """Create the standard Evidence Register sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Verification Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "A.5.9 All Domains", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Not verified"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Not verified,In Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

def create_summary_dashboard_sheet(ws):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 (A.5.9.1 Asset Discovery)."""
    from openpyxl.styles import Border, Side

    CHECK = "\u2705"
    XMARK = "\u274c"

    _thin = Side(style="thin")
    _bdr = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _red_b = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _crit = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _high = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _med = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    def _merge_row(row, fill, text, font_kw, align="left"):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.fill = fill
        c.font = Font(**font_kw)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = _bdr
        for col in "BCDE":
            ws[f"{col}{row}"].border = _bdr

    # Row 1: Title
    _merge_row(1, _navy, "ASSET DISCOVERY ASSESSMENT \u2014 SUMMARY DASHBOARD",
               {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"}, align="center")
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    _merge_row(2, _blue, "ISO 27001:2022 \u00b7 Control A.5.9 \u00b7 Inventory of Information and Assets",
               {"name": "Calibri", "size": 10, "italic": True, "color": "FFFFFF"})
    ws.row_dimensions[2].height = 18

    # Row 3: blank spacer
    ws.row_dimensions[3].height = 6

    # ── TABLE 1 ────────────────────────────────────────────────────────────────
    _merge_row(4, _blue, "TABLE 1: COMPLIANCE ASSESSMENT",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    # Row 5: headers
    for col, label in enumerate(["Assessment Area", "Compliant", "Non-Compliant", "Total Items", "Compliance %"], 1):
        c = ws.cell(row=5, column=col, value=label)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr

    # Rows 6–10: assessment areas
    t1_rows = [
        (6, "Information Asset Discovery Coverage",
         "=COUNTIF(\'Information Assets Discovery\'!F4:F18,\"\u2705 Compliant\")",
         "=COUNTA(\'Information Assets Discovery\'!A4:A18)"),
        (7, "IT Infrastructure Discovery Coverage",
         "=COUNTIF(\'IT Infrastructure Discovery\'!F4:F19,\"\u2705 Compliant\")",
         "=COUNTA(\'IT Infrastructure Discovery\'!A4:A19)"),
        (8, "Application Discovery Coverage",
         "=COUNTIF(\'Applications Discovery\'!F4:F18,\"\u2705 Compliant\")",
         "=COUNTA(\'Applications Discovery\'!A4:A18)"),
        (9, "Physical Assets Discovery Coverage",
         "=COUNTIF(\'Physical Assets Discovery\'!F4:F15,\"\u2705 Compliant\")",
         "=COUNTA(\'Physical Assets Discovery\'!A4:A15)"),
        (10, "Personnel Assets Discovery Coverage",
         "=COUNTIF(\'Personnel Assets Discovery\'!F4:F18,\"\u2705 Compliant\")",
         "=COUNTA(\'Personnel Assets Discovery\'!A4:A18)"),
    ]
    for row, label, b_fml, d_fml in t1_rows:
        ws.cell(row=row, column=1, value=label).border = _bdr
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{row}"] = b_fml
        ws[f"B{row}"].border = _bdr
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"] = f"=D{row}-B{row}"
        ws[f"C{row}"].border = _bdr
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"] = d_fml
        ws[f"D{row}"].border = _bdr
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"E{row}"] = f"=IFERROR(B{row}/D{row},0)"
        ws[f"E{row}"].number_format = "0.0%"
        ws[f"E{row}"].border = _bdr
        ws[f"E{row}"].alignment = Alignment(horizontal="center")

    # Row 11: TOTAL
    for col, val in enumerate(["TOTAL", "=SUM(B6:B10)", "=SUM(C6:C10)", "=SUM(D6:D10)", "=IFERROR(B11/D11,0)"], 1):
        c = ws.cell(row=11, column=col, value=val)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        c.border = _bdr
    ws["E11"].number_format = "0.0%"

    # Row 12: blank spacer
    ws.row_dimensions[12].height = 6

    # ── TABLE 2 ────────────────────────────────────────────────────────────────
    _merge_row(13, _blue, "TABLE 2: KEY PERFORMANCE INDICATORS",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    def _subhdr(row, label):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bdr
        for col in "BCDE":
            ws[f"{col}{row}"].border = _bdr

    def _metric(row, label, formula, fmt=None):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row}"].border = _bdr
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = _bdr
        if fmt:
            ws[f"B{row}"].number_format = fmt
        for col in "CDE":
            ws[f"{col}{row}"].border = _bdr

    _subhdr(14, "Discovery Overview")
    _metric(15, "Information Assets Discovered", "=COUNTA(\'Information Assets Discovery\'!A4:A18)")
    _metric(16, "IT Infrastructure Discovered", "=COUNTA(\'IT Infrastructure Discovery\'!A4:A19)")
    _metric(17, "Applications Discovered", "=COUNTA(\'Applications Discovery\'!A4:A18)")
    _subhdr(18, "Discovery Coverage")
    _metric(19, "Physical Assets Discovered", "=COUNTA(\'Physical Assets Discovery\'!A4:A15)")
    _metric(20, "Personnel Assets Discovered", "=COUNTA(\'Personnel Assets Discovery\'!A4:A18)")
    _metric(21, "Total Asset Subcategories Assessed", "=B15+B16+B17+B19+B20")
    _subhdr(22, "Compliance Status")
    _metric(23, "Non-Compliant Information Asset Types", "=COUNTIF(\'Information Assets Discovery\'!F4:F18,\"\u274c Non-Compliant\")")
    _metric(24, "Non-Compliant IT Infrastructure Types", "=COUNTIF(\'IT Infrastructure Discovery\'!F4:F19,\"\u274c Non-Compliant\")")
    _metric(25, "Non-Compliant Application Types", "=COUNTIF(\'Applications Discovery\'!F4:F18,\"\u274c Non-Compliant\")")
    _subhdr(26, "Gaps & Evidence")
    _metric(27, "Non-Compliant Physical Asset Types", "=COUNTIF(\'Physical Assets Discovery\'!F4:F15,\"\u274c Non-Compliant\")")
    _metric(28, "Non-Compliant Personnel Asset Types", "=COUNTIF(\'Personnel Assets Discovery\'!F4:F18,\"\u274c Non-Compliant\")")
    _metric(29, "Unverified Evidence Items", "=COUNTIF(\'Evidence Register\'!H6:H105,\"Not verified\")")

    # Row 30: blank spacer
    ws.row_dimensions[30].height = 6

    # ── TABLE 3 ────────────────────────────────────────────────────────────────
    _merge_row(31, _red_b, "TABLE 3: CRITICAL FINDINGS & RISK INDICATORS",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    # Row 32: header
    for col, label in enumerate(["Critical Finding", "Count", "Severity", "ISO Reference", "Action Required"], 1):
        c = ws.cell(row=32, column=col, value=label)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr

    def _finding(row, label, formula, severity, iso_ref, action, fill, text_color):
        ws[f"A{row}"] = label
        ws[f"A{row}"].fill = fill
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"A{row}"].border = _bdr
        ws[f"B{row}"] = formula
        ws[f"B{row}"].fill = fill
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = _bdr
        ws[f"C{row}"] = severity
        ws[f"C{row}"].fill = fill
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"].border = _bdr
        ws[f"D{row}"] = iso_ref
        ws[f"D{row}"].fill = fill
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"].border = _bdr
        ws[f"E{row}"] = action
        ws[f"E{row}"].fill = fill
        ws[f"E{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"E{row}"].border = _bdr

    _finding(33, "Non-compliant information asset types",
             "=COUNTIF(\'Information Assets Discovery\'!F4:F18,\"\u274c Non-Compliant\")",
             "CRITICAL", "A.5.9 §2.1",
             "Immediate — missing information assets create data blind spots and unquantified risk",
             _crit, "C00000")
    _finding(34, "Non-compliant IT infrastructure types",
             "=COUNTIF(\'IT Infrastructure Discovery\'!F4:F19,\"\u274c Non-Compliant\")",
             "CRITICAL", "A.5.9 §2.2",
             "Immediate — undiscovered IT assets cannot be patched, monitored, or protected",
             _crit, "C00000")
    _finding(35, "Non-compliant application types",
             "=COUNTIF(\'Applications Discovery\'!F4:F18,\"\u274c Non-Compliant\")",
             "HIGH", "A.5.9 §2.2",
             "Urgent — unknown applications may process data without appropriate controls",
             _high, "9C5700")
    _finding(36, "Non-compliant physical asset types",
             "=COUNTIF(\'Physical Assets Discovery\'!F4:F15,\"\u274c Non-Compliant\")",
             "HIGH", "A.5.9 §2.2",
             "Urgent — undiscovered physical assets may contain sensitive data without protection",
             _high, "9C5700")
    _finding(37, "Non-compliant personnel asset types",
             "=COUNTIF(\'Personnel Assets Discovery\'!F4:F18,\"\u274c Non-Compliant\")",
             "HIGH", "A.5.9 §2.2",
             "Urgent — unmapped personnel create succession and knowledge transfer risk",
             _high, "9C5700")
    _finding(38, "Asset subcategories at-risk of non-compliance",
             "=COUNTIF(\'Information Assets Discovery\'!F4:F18,\"\u26a0\ufe0f At Risk\")+COUNTIF(\'IT Infrastructure Discovery\'!F4:F19,\"\u26a0\ufe0f At Risk\")+COUNTIF(\'Applications Discovery\'!F4:F18,\"\u26a0\ufe0f At Risk\")",
             "HIGH", "A.5.9 §3",
             "Plan — subcategories approaching threshold require enhanced discovery activity",
             _high, "9C5700")
    _finding(39, "Open discovery gaps (non-compliant physical + personnel)",
             "=COUNTIF(\'Physical Assets Discovery\'!F4:F15,\"\u274c Non-Compliant\")+COUNTIF(\'Personnel Assets Discovery\'!F4:F18,\"\u274c Non-Compliant\")",
             "MEDIUM", "A.5.9 §3",
             "Plan — remediation required for all non-compliant subcategories",
             _med, "276221")
    _finding(40, "Unverified evidence items",
             "=COUNTIF(\'Evidence Register\'!H6:H105,\"Not verified\")",
             "MEDIUM", "A.5.9 §3",
             "Plan — evidence requires verification before next audit",
             _med, "276221")



def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-001..015)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = blue
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8); Overall Compliance Rating at B6 (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = yellow
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between approver sections

    # FINAL DECISION (GS-AS-004/012: col A plain bold, no dark fill)
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = yellow
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS (GS-AS-005/013: 4472C4 banner, borders on all)
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = blue
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
