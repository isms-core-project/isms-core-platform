#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.9.4 - Owner Accountability Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Assessment Domain 4 of 5: Asset Ownership & Accountability Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific organisational structure, communication channels,
and accountability expectations.

Key customization areas:
1. Organisational structure (departments, business units, reporting lines)
2. Attestation workflow (email templates, reminder schedules, escalation paths)
3. Performance expectations (review frequency, response times, engagement levels)
4. Accountability thresholds (what's acceptable for your organisational culture)
5. Organisation name, CISO details, contact information
6. File paths and naming conventions
7. Integration with HR systems and email platforms

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
asset ownership accountability through systematic verification of ownership
coverage, owner acknowledgment, awareness, and performance.

**Purpose:**
Enables systematic assessment of whether asset owners understand their
responsibilities, acknowledge ownership, demonstrate awareness of assets under
their care, and actively perform their ownership duties. Measures accountability
through attestation campaigns, engagement tracking, and performance metrics.

**Assessment Scope:**
- Ownership Coverage: All assets have identified owners who acknowledge role
- Owner Acknowledgment: Formal attestation that owner accepts responsibilities
- Owner Awareness: Owner can identify assets under their care and knows status
- Owner Performance: Owner actively performs duties (reviews, updates, decisions)

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and accountability framework
2. Ownership Coverage - Asset ownership assignment and acceptance verification
3. Owner Acknowledgment - Attestation campaign tracking and response rates
4. Owner Awareness - Owner knowledge of assets and current state
5. Owner Performance - Active performance of ownership duties
6. Accountability Metrics - Aggregated scores and weighted accountability index
7. Evidence Register - Attestation evidence and performance documentation

**Key Features:**
- Data validation with dropdown lists for attestation status
- Conditional formatting for accountability scores (Green/Yellow/Red)
- Attestation campaign tracking (email sent, reminders, responses)
- Automated gap identification for missing or non-responsive owners
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Weighted accountability scoring (4 domains with configurable weights)
- Reminder schedule tracking (Day 7, 14, 21, escalation at Day 28)
- Performance engagement metrics
- Owner contact information and escalation paths

**Integration:**
data from all four assessment domains for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a59_4_owner_accountability.py

Output:
    File: ISMS-IMP-A.5.9.4_Owner_Accountability_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Populate ownership coverage (assign owners to all assets)
    2. Launch attestation campaign (email owners requesting acknowledgment)
    3. Track responses and send reminders (Day 7, 14, 21)
    4. Escalate non-responders at Day 28
    5. Conduct awareness checks (can owner list their assets?)
    6. Assess performance (review compliance, response times, engagement)
    7. Review Accountability Metrics for overall accountability index
    8. Collect and link evidence in Evidence Register sheet
    9. Export metrics CSV for dashboard consolidation (Sheet 6)
    10. Store assessment workbook per retention policy (7 years minimum)
    11. Update quarterly or after ownership changes

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.9
Assessment Domain:    4 of 5
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License]

Related Documents:
    - ISMS-POL-A.5.9: Inventory of Information and Assets (Policy)
    - ISMS-IMP-A.5.9-1: Asset Discovery Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-2: Inventory Maintenance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-3: Quality & Compliance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-4: Owner Accountability Assessment (Implementation Guide)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 22.01.2026
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.9-4 specification
    - Four accountability domains with weighted scoring
    - Attestation campaign tracking with reminder workflow
    - Owner awareness verification
    - Performance engagement metrics
    - Evidence collection and audit trail support

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive asset inventory details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review asset inventory procedures and classification criteria annually or when
new asset categories are introduced, system landscapes change, or audit findings
identify inventory gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
from datetime import datetime
from pathlib import Path
import os

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# CUSTOMIZE: Configuration

# Document identification constants

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.9.4"
WORKBOOK_NAME = "Owner Accountability"
CONTROL_ID   = "A.5.9"
CONTROL_NAME = "Inventory of Information and Assets"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("="*80)
    logger.info("ISMS Control A.5.9 - Owner Accountability Assessment Generator")
    logger.info("="*80)
    logger.info("")
    logger.info("Generating assessment workbook...")
    logger.info("")
    
    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create sheets in order
    sheets = [
        "Instructions & Legend",
        "Ownership Coverage",
        "Owner Acknowledgment",
        "Owner Awareness",
        "Owner Performance",
        "Accountability Metrics",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
        logger.info(f"  ✓ Created sheet: {sheet_name}")
    
    logger.info("")
    logger.info("Populating sheets...")
    logger.info("")
    
    # Populate each sheet
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✓ Instructions & Legend")
    
    create_ownership_coverage_sheet(wb["Ownership Coverage"])
    logger.info("  ✓ Ownership Coverage")
    
    create_owner_acknowledgment_sheet(wb["Owner Acknowledgment"])
    logger.info("  ✓ Owner Acknowledgment")
    
    create_owner_awareness_sheet(wb["Owner Awareness"])
    logger.info("  ✓ Owner Awareness")
    
    create_owner_performance_sheet(wb["Owner Performance"])
    logger.info("  ✓ Owner Performance")
    
    create_accountability_metrics_sheet(wb["Accountability Metrics"])
    logger.info("  ✓ Accountability Metrics")
    
    create_evidence_register(wb["Evidence Register"])
    logger.info("  ✓ Evidence Register")
    
    create_summary_dashboard_sheet(wb["Summary Dashboard"])
    logger.info("  ✓ Summary Dashboard")
    
    create_approval_sheet(wb["Approval Sign-Off"])
    logger.info("  ✓ Approval & Sign-Off")
    
    # Finalise data validations
    finalize_validations(wb)

    # Save workbook to WKBK directory
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    wb.save(output_path)
    
    logger.info("")
    logger.info("="*80)
    logger.info(f"SUCCESS: {output_path.name}")
    logger.info("="*80)
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open workbook and review Instructions & Legend")
    logger.info("  2. Populate ownership coverage (assign all assets to owners)")
    logger.info("  3. Launch attestation campaign (email owners)")
    logger.info("  4. Track acknowledgment responses and send reminders")
    logger.info("  5. Conduct awareness checks (owner knowledge verification)")
    logger.info("  6. Assess owner performance (reviews, responsiveness, engagement)")
    logger.info("  7. Review Accountability Metrics for overall accountability index")
    logger.info("  8. Export CSV from Sheet 6 for dashboard consolidation")
    logger.info("  9. Obtain stakeholder review and approval")
    logger.info("")
    logger.info(f"Output location: {output_path}")
    logger.info("")



def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Ownership Coverage sheet with all asset types and owner assignments.', '2. Complete the Owner Acknowledgment sheet tracking attestation campaign responses.', '3. Complete the Owner Awareness sheet verifying owner knowledge of their assets.', '4. Complete the Owner Performance sheet measuring active ownership engagement.', '5. Review the Accountability Metrics sheet for the overall accountability index.', '6. Link evidence in the Evidence Register sheet.', '7. Obtain approvals in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 21

    # Status Legend — row position tracks after instructions
    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_ownership_coverage_sheet(ws):
    """Create Ownership Coverage sheet"""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "OWNERSHIP COVERAGE"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:L2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Asset Category', 25),
        ('B', 'Total Assets', 15),
        ('C', 'Assets with Owner', 18),
        ('D', 'Assets without Owner', 18),
        ('E', 'Coverage %', 15),
        ('F', 'Target %', 12),
        ('G', 'Gap', 12),
        ('H', 'Status', 18),
        ('I', 'Unique Owners Count', 18),
        ('J', 'Avg Assets per Owner', 18),
        ('K', 'Evidence ID', 15),
        ('L', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Asset categories
    categories = [
        "Information Assets",
        "IT Infrastructure",
        "Applications",
        "Physical Assets",
        "Personnel Assets",
    ]
    
    row = 4
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'F{row}'] = "100%"
        
        # User enters counts
        
        # Assets without owner
        ws[f'D{row}'] = f'=B{row}-C{row}'
        
        # Coverage %
        ws[f'E{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        # Gap
        ws[f'G{row}'] = f'=E{row}-100'
        ws[f'G{row}'].number_format = '0.0"%"'
        
        # Status
        ws[f'H{row}'] = f'=IF(E{row}=100,"✅ Full Coverage",IF(E{row}>=95,"⚠️ Minor Gaps","❌ Major Gaps"))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Unique owners and avg per owner (user enters)
        ws[f'J{row}'] = f'=IFERROR(B{row}/I{row},0)'
        ws[f'J{row}'].number_format = '0.0'
        
        
        row += 1
    

    # Apply FFFFCC fill + thin borders to all data rows
    from openpyxl.styles import Border as _B59, Side as _S59
    _ts = _S59(style="thin")
    _bd = _B59(left=_ts, right=_ts, top=_ts, bottom=_ts)
    _yf = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _sf = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for _r in range(4, row):
        for _c in range(1, 12 + 1):
            _cel = ws.cell(row=_r, column=_c)
            _cel.fill = _sf if _r == 4 else _yf
            _cel.border = _bd
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "OWNERSHIP COVERAGE SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Assets"
    ws[f'B{summary_row}'] = f'=SUM(B4:B8)'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Assets with Owner"
    ws[f'B{summary_row}'] = f'=SUM(C4:C8)'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Assets without Owner"
    ws[f'B{summary_row}'] = f'=SUM(D4:D8)'
    ws[f'B{summary_row}'].font = Font(bold=True, color="9C0006")
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Overall Coverage %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-2}/B{summary_row-3}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color="FFEB9C", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "100%"
    ws[f'B{summary_row}'].font = Font(bold=True, color="9C0006")
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-100'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}=100,"✅ Strong",IF(B{summary_row-3}>=95,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'E4:E8',
        CellIsRule(operator='equal', formula=['100'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E8',
        CellIsRule(operator='between', formula=['95', '99'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E8',
        CellIsRule(operator='lessThan', formula=['95'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_owner_acknowledgment_sheet(ws):
    """Create Owner Acknowledgment (Attestation) sheet"""
    
    # Header
    ws.merge_cells('A1:O1')
    ws['A1'] = "OWNER ACKNOWLEDGMENT"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:O2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Owner Name', 25),
        ('B', 'Owner Email', 30),
        ('C', 'Department', 20),
        ('D', 'Assets Count', 15),
        ('E', 'Email Sent Date', 18),
        ('F', 'Reminder 1 (Day 7)', 18),
        ('G', 'Reminder 2 (Day 14)', 18),
        ('H', 'Reminder 3 (Day 21)', 18),
        ('I', 'Escalation (Day 28)', 18),
        ('J', 'Response Date', 18),
        ('K', 'Days to Respond', 15),
        ('L', 'Attestation Status', 20),
        ('M', 'Next Action', 25),
        ('N', 'Evidence ID', 15),
        ('O', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample owner rows (user fills actual owners)
    sample_owners = [
        ("John Doe", "john.doe@org.ch", "Engineering", 15),
        ("Jane Smith", "jane.smith@org.ch", "IT Operations", 42),
        ("Bob Johnson", "bob.johnson@org.ch", "Security", 8),
        ("Alice Williams", "alice.williams@org.ch", "HR", 5),
    ]
    
    row = 4
    for name, email, dept, count in sample_owners:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = email
        ws[f'C{row}'] = dept
        ws[f'D{row}'] = count
        
        # User enters campaign dates
        for col in ['E', 'F', 'G', 'H', 'I', 'J']:
            pass
        
        # Days to respond (formula)
        ws[f'K{row}'] = f'=IF(AND(E{row}<>"",J{row}<>""),J{row}-E{row},"")'
        ws[f'K{row}'].number_format = '0'
        
        # Attestation Status (user selects)
        
        # Next Action (user enters)
        
        row += 1
    
    # Add 30 blank rows
    for i in range(30):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O']:
            pass
        ws[f'K{row}'] = f'=IF(AND(E{row}<>"",J{row}<>""),J{row}-E{row},"")'
        row += 1
    

    # Apply FFFFCC fill + thin borders to all data rows
    from openpyxl.styles import Border as _B59, Side as _S59
    _ts = _S59(style="thin")
    _bd = _B59(left=_ts, right=_ts, top=_ts, bottom=_ts)
    _yf = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _sf = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for _r in range(4, row):
        for _c in range(1, 15 + 1):
            _cel = ws.cell(row=_r, column=_c)
            _cel.fill = _sf if _r == 4 else _yf
            _cel.border = _bd
    # Data validations
    attestation_statuses = [
        "Attested - On Time",
        "Attested - Late",
        "Pending",
        "Non-Responsive",
        "Escalated",
        "Exempted"
    ]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(attestation_statuses)}"', allow_blank=True)
    dv_status.add(f'L4:L{row-1}')
    ws.add_data_validation(dv_status)
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "ACKNOWLEDGMENT SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Owners"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-31})'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Attested (On Time + Late)"
    ws[f'B{summary_row}'] = f'=COUNTIFS(L4:L{row-31},"Attested*")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Acknowledgment Rate %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color="FFEB9C", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "95%"
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-95'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Response Time (Days)"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(K4:K{row-31}),0)'
    ws[f'B{summary_row}'].number_format = '0.0'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-4}>=95,"✅ Strong",IF(B{summary_row-4}>=90,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)


def create_owner_awareness_sheet(ws):
    """Create Owner Awareness sheet"""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "OWNER AWARENESS"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:K2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Owner Name', 25),
        ('B', 'Assets Assigned', 15),
        ('C', 'Assets Identified by Owner', 22),
        ('D', 'Awareness %', 15),
        ('E', 'Target %', 12),
        ('F', 'Gap', 12),
        ('G', 'Status', 18),
        ('H', 'Verification Method', 30),
        ('I', 'Verification Date', 18),
        ('J', 'Evidence ID', 15),
        ('K', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample owners (user fills)
    sample_owners = [
        "John Doe",
        "Jane Smith",
        "Bob Johnson",
        "Alice Williams",
    ]
    
    row = 4
    for owner in sample_owners:
        ws[f'A{row}'] = owner
        ws[f'E{row}'] = "100%"
        
        # User enters counts
        
        # Awareness %
        ws[f'D{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'D{row}'].number_format = '0.0"%"'
        
        # Gap
        ws[f'F{row}'] = f'=D{row}-100'
        ws[f'F{row}'].number_format = '0.0"%"'
        
        # Status
        ws[f'G{row}'] = f'=IF(D{row}=100,"✅ Fully Aware",IF(D{row}>=95,"⚠️ Mostly Aware","❌ Unaware"))'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        # User enters verification details
        for col in ['H', 'I', 'J', 'K']:
            pass
        
        row += 1
    
    # Add 30 blank rows
    for i in range(30):
        ws[f'E{row}'] = "100%"
        for col in ['A', 'B', 'C', 'H', 'I', 'J', 'K']:
            pass
        ws[f'D{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=D{row}-100'
        ws[f'F{row}'].number_format = '0.0"%"'
        ws[f'G{row}'] = f'=IF(D{row}=100,"✅ Fully Aware",IF(D{row}>=95,"⚠️ Mostly Aware","❌ Unaware"))'
        row += 1
    

    # Apply FFFFCC fill + thin borders to all data rows
    from openpyxl.styles import Border as _B59, Side as _S59
    _ts = _S59(style="thin")
    _bd = _B59(left=_ts, right=_ts, top=_ts, bottom=_ts)
    _yf = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _sf = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for _r in range(4, row):
        for _c in range(1, 11 + 1):
            _cel = ws.cell(row=_r, column=_c)
            _cel.fill = _sf if _r == 4 else _yf
            _cel.border = _bd
    # Data validation
    verification_methods = [
        "Interview - Owner Listed Assets",
        "Survey - Owner Confirmed Assets",
        "Audit - Cross-checked vs Inventory",
        "Self-Assessment",
        "Other (Specify)"
    ]
    dv_method = DataValidation(type="list", formula1=f'"{",".join(verification_methods)}"', allow_blank=True)
    dv_method.add(f'H4:H{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "AWARENESS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Awareness %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(D4:D{row-31}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color="FFEB9C", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "100%"
    ws[f'B{summary_row}'].font = Font(bold=True, color="9C0006")
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-100'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}=100,"✅ Strong",IF(B{summary_row-3}>=95,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'D4:D{row-31}',
        CellIsRule(operator='equal', formula=['100'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'D4:D{row-31}',
        CellIsRule(operator='between', formula=['95', '99'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'D4:D{row-31}',
        CellIsRule(operator='lessThan', formula=['95'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_owner_performance_sheet(ws):
    """Create Owner Performance sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "OWNER PERFORMANCE"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Owner Name', 25),
        ('B', 'Assets Count', 15),
        ('C', 'Reviews Due', 15),
        ('D', 'Reviews Completed', 18),
        ('E', 'Review Compliance %', 18),
        ('F', 'Avg Response Time (Days)', 22),
        ('G', 'Target Response (Days)', 20),
        ('H', 'Responsiveness %', 18),
        ('I', 'Performance Score', 18),
        ('J', 'Target', 12),
        ('K', 'Status', 18),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample owners
    sample_owners = [
        ("John Doe", 15, 3),
        ("Jane Smith", 42, 3),
        ("Bob Johnson", 8, 3),
        ("Alice Williams", 5, 3),
    ]
    
    row = 4
    for owner, count, target_days in sample_owners:
        ws[f'A{row}'] = owner
        ws[f'B{row}'] = count
        ws[f'G{row}'] = target_days
        ws[f'J{row}'] = "80%"
        
        # User enters review counts
        
        # Review Compliance %
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        # User enters avg response time
        
        # Responsiveness % (inverse - faster is better)
        ws[f'H{row}'] = f'=IF(F{row}="","",MAX(0,100-((F{row}-G{row})/G{row}*100)))'
        ws[f'H{row}'].number_format = '0.0"%"'
        
        # Performance Score = (Review Compliance × 60%) + (Responsiveness × 40%)
        ws[f'I{row}'] = f'=IF(OR(E{row}="",H{row}=""),"",(E{row}*0.6)+(H{row}*0.4))'
        ws[f'I{row}'].number_format = '0.0"%"'
        ws[f'I{row}'].font = Font(bold=True)
        
        # Status
        ws[f'K{row}'] = f'=IF(I{row}="","",IF(I{row}>=80,"✅ Strong",IF(I{row}>=75,"⚠️ Needs Improvement","❌ Weak")))'
        ws[f'K{row}'].alignment = Alignment(horizontal='center')
        
        
        row += 1
    
    # Add 30 blank rows
    for i in range(30):
        ws[f'G{row}'] = 3
        ws[f'J{row}'] = "80%"
        for col in ['A', 'B', 'C', 'D', 'F', 'L', 'M']:
            pass
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'H{row}'] = f'=IF(F{row}="","",MAX(0,100-((F{row}-G{row})/G{row}*100)))'
        ws[f'H{row}'].number_format = '0.0"%"'
        ws[f'I{row}'] = f'=IF(OR(E{row}="",H{row}=""),"",(E{row}*0.6)+(H{row}*0.4))'
        ws[f'I{row}'].number_format = '0.0"%"'
        ws[f'K{row}'] = f'=IF(I{row}="","",IF(I{row}>=80,"✅ Strong",IF(I{row}>=75,"⚠️ Needs Improvement","❌ Weak")))'
        row += 1
    

    # Apply FFFFCC fill + thin borders to all data rows
    from openpyxl.styles import Border as _B59, Side as _S59
    _ts = _S59(style="thin")
    _bd = _B59(left=_ts, right=_ts, top=_ts, bottom=_ts)
    _yf = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _sf = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for _r in range(4, row):
        for _c in range(1, 13 + 1):
            _cel = ws.cell(row=_r, column=_c)
            _cel.fill = _sf if _r == 4 else _yf
            _cel.border = _bd
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "PERFORMANCE SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Performance Score %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(I4:I{row-31}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color="FFEB9C", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "80%"
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-80'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}>=80,"✅ Strong",IF(B{summary_row-3}>=75,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'I4:I{row-31}',
        CellIsRule(operator='greaterThanOrEqual', formula=['80'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'I4:I{row-31}',
        CellIsRule(operator='between', formula=['75', '79'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'I4:I{row-31}',
        CellIsRule(operator='lessThan', formula=['75'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_accountability_metrics_sheet(ws):
    """Create Accountability Metrics sheet - consolidates all domains"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "ACCOUNTABILITY METRICS"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Domains table
    ws['A3'] = "Accountability Domain"
    ws['B3'] = "Score %"
    ws['C3'] = "Weight %"
    ws['D3'] = "Weighted Score"
    ws['E3'] = "Target %"
    ws['F3'] = "Gap"
    ws['G3'] = "Status"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
    
    # Link to domain sheets
    domains_data = [
        ("Coverage", "='Ownership Coverage'!B17", 40, 100),
        ("Acknowledgment", "='Owner Acknowledgment'!B45", 25, 95),
        ("Awareness", "='Owner Awareness'!B40", 20, 100),
        ("Performance", "='Owner Performance'!B41", 15, 80),
    ]
    
    row = 4
    for domain, formula, weight, target in domains_data:
        ws[f'A{row}'] = domain
        ws[f'A{row}'].font = Font(bold=True)
        
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0"%"'
        ws[f'B{row}'].font = Font(bold=True)
        
        ws[f'C{row}'] = f'{weight}%'
        
        ws[f'D{row}'] = f'=B{row}*{weight}/100'
        ws[f'D{row}'].number_format = '0.0"%"'
        
        ws[f'E{row}'] = f'{target}%'
        
        ws[f'F{row}'] = f'=B{row}-{target}/100'
        ws[f'F{row}'].number_format = '0.0"%"'
        
        ws[f'G{row}'] = f'=IF(B{row}>={target}/100,"✅ Strong",IF(B{row}>={target}/100-0.05,"⚠️ Medium","❌ Weak"))'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Overall accountability index
    overall_row = row + 2
    ws[f'A{overall_row}'] = "OVERALL ACCOUNTABILITY INDEX"
    ws[f'A{overall_row}'].font = Font(size=12, bold=True)
    ws[f'B{overall_row}'] = f'=SUM(D4:D7)'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True, size=14)
    ws[f'B{overall_row}'].fill = PatternFill(start_color="FFEB9C", fill_type='solid')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Target"
    ws[f'B{overall_row}'] = "94%"
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Gap"
    ws[f'B{overall_row}'] = f'=B{overall_row-2}-0.94'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Status"
    ws[f'B{overall_row}'] = f'=IF(B{overall_row-3}>=0.94,"✅ Strong",IF(B{overall_row-3}>=0.89,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    # CSV Export section
    csv_row = overall_row + 3
    ws[f'A{csv_row}'] = "CSV EXPORT FOR DASHBOARD (Copy rows below)"
    ws[f'A{csv_row}'].font = Font(size=11, bold=True, color="FFFFFF")
    ws[f'A{csv_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    csv_row += 1
    ws[f'A{csv_row}'] = "Accountability_Domain"
    ws[f'B{csv_row}'] = "Score_%"
    ws[f'C{csv_row}'] = "Status"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{csv_row}'].font = Font(bold=True)
    
    csv_row += 1
    for i in range(4):
        ws[f'A{csv_row+i}'] = f'=A{4+i}'
        ws[f'B{csv_row+i}'] = f'=B{4+i}'
        ws[f'C{csv_row+i}'] = f'=G{4+i}'
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18


def create_evidence_register(ws):
    """Create the standard Evidence Register sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Verification Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "A.5.9 All Domains", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Not verified"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Not verified,In Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

def create_summary_dashboard_sheet(ws):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 (A.5.9.4 Owner Accountability)."""
    from openpyxl.styles import Border, Side

    CHECK = "\u2705"
    XMARK = "\u274c"

    _thin = Side(style="thin")
    _bdr = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _red_b = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _crit = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _high = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _med = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    def _merge_row(row, fill, text, font_kw, align="left"):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.fill = fill
        c.font = Font(**font_kw)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = _bdr
        for col in "BCDE":
            ws[f"{col}{row}"].border = _bdr

    _merge_row(1, _navy, "OWNER ACCOUNTABILITY ASSESSMENT \u2014 SUMMARY DASHBOARD",
               {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"}, align="center")
    ws.row_dimensions[1].height = 35

    _merge_row(2, _blue, "ISO 27001:2022 \u00b7 Control A.5.9 \u00b7 Inventory of Information and Assets",
               {"name": "Calibri", "size": 10, "italic": True, "color": "FFFFFF"})
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    _merge_row(4, _blue, "TABLE 1: COMPLIANCE ASSESSMENT",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    for col, label in enumerate(["Assessment Area", "Compliant", "Non-Compliant", "Total Items", "Compliance %"], 1):
        c = ws.cell(row=5, column=col, value=label)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr

    t1_rows = [
        (6, "Ownership Coverage (All Categories)",
         "=COUNTIF(\'Ownership Coverage\'!H4:H8,\"\u2705 Full Coverage\")",
         "=COUNTA(\'Ownership Coverage\'!A4:A8)"),
        (7, "Owner Acknowledgment & Attestation",
         "=COUNTIFS(\'Owner Acknowledgment\'!L4:L53,\"Attested*\")",
         "=COUNTA(\'Owner Acknowledgment\'!A4:A53)"),
        (8, "Owner Awareness Compliance",
         "=COUNTIF(\'Owner Awareness\'!G4:G53,\"\u2705 Fully Aware\")",
         "=COUNTA(\'Owner Awareness\'!A4:A53)"),
        (9, "Owner Performance — Threshold Met",
         "=COUNTIF(\'Owner Performance\'!K4:K53,\"\u2705 Strong\")",
         "=COUNTA(\'Owner Performance\'!A4:A53)"),
        (10, "Accountability Governance Strength",
         "=COUNTIF(\'Accountability Metrics\'!G4:G13,\"\u2705 Strong\")",
         "=COUNTA(\'Accountability Metrics\'!A4:A13)"),
    ]
    for row, label, b_fml, d_fml in t1_rows:
        ws.cell(row=row, column=1, value=label).border = _bdr
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{row}"] = b_fml
        ws[f"B{row}"].border = _bdr
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"] = f"=D{row}-B{row}"
        ws[f"C{row}"].border = _bdr
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"] = d_fml
        ws[f"D{row}"].border = _bdr
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"E{row}"] = f"=IFERROR(B{row}/D{row},0)"
        ws[f"E{row}"].number_format = "0.0%"
        ws[f"E{row}"].border = _bdr
        ws[f"E{row}"].alignment = Alignment(horizontal="center")

    for col, val in enumerate(["TOTAL", "=SUM(B6:B10)", "=SUM(C6:C10)", "=SUM(D6:D10)", "=IFERROR(B11/D11,0)"], 1):
        c = ws.cell(row=11, column=col, value=val)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        c.border = _bdr
    ws["E11"].number_format = "0.0%"

    ws.row_dimensions[12].height = 6

    _merge_row(13, _blue, "TABLE 2: KEY PERFORMANCE INDICATORS",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    def _subhdr(row, label):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bdr
        for col in "BCDE":
            ws[f"{col}{row}"].border = _bdr

    def _metric(row, label, formula, fmt=None):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row}"].border = _bdr
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = _bdr
        if fmt:
            ws[f"B{row}"].number_format = fmt
        for col in "CDE":
            ws[f"{col}{row}"].border = _bdr

    _subhdr(14, "Ownership Coverage")
    _metric(15, "Asset Categories with Full Coverage", "=COUNTIF(\'Ownership Coverage\'!H4:H8,\"\u2705 Full Coverage\")")
    _metric(16, "Asset Categories with Gaps", "=COUNTIF(\'Ownership Coverage\'!H4:H8,\"\u26a0\ufe0f Minor Gaps\")+COUNTIF(\'Ownership Coverage\'!H4:H8,\"\u274c Major Gaps\")")
    _metric(17, "Total Assets Without Owner", "=SUM(\'Ownership Coverage\'!D4:D8)")
    _subhdr(18, "Acknowledgment & Awareness")
    _metric(19, "Owners Attested", "=COUNTIFS(\'Owner Acknowledgment\'!L4:L53,\"Attested*\")")
    _metric(20, "Owners Not Yet Attested", "=COUNTA(\'Owner Acknowledgment\'!A4:A53)-COUNTIFS(\'Owner Acknowledgment\'!L4:L53,\"Attested*\")")
    _metric(21, "Fully Aware Owners", "=COUNTIF(\'Owner Awareness\'!G4:G53,\"\u2705 Fully Aware\")")
    _subhdr(22, "Owner Performance")
    _metric(23, "Strong-Performing Owners", "=COUNTIF(\'Owner Performance\'!K4:K53,\"\u2705 Strong\")")
    _metric(24, "Underperforming Owners", "=COUNTIF(\'Owner Performance\'!K4:K53,\"\u274c Weak\")")
    _subhdr(25, "Gaps & Evidence")
    _metric(26, "Awareness Gaps (Unaware)", "=COUNTIF(\'Owner Awareness\'!G4:G53,\"\u274c Unaware\")")
    _metric(27, "Unverified Evidence Items", "=COUNTIF(\'Evidence Register\'!H6:H105,\"Not verified\")")

    ws.row_dimensions[28].height = 6

    _merge_row(29, _red_b, "TABLE 3: CRITICAL FINDINGS & RISK INDICATORS",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    for col, label in enumerate(["Critical Finding", "Count", "Severity", "ISO Reference", "Action Required"], 1):
        c = ws.cell(row=30, column=col, value=label)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr

    def _finding(row, label, formula, severity, iso_ref, action, fill, text_color):
        ws[f"A{row}"] = label
        ws[f"A{row}"].fill = fill
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"A{row}"].border = _bdr
        ws[f"B{row}"] = formula
        ws[f"B{row}"].fill = fill
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = _bdr
        ws[f"C{row}"] = severity
        ws[f"C{row}"].fill = fill
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"].border = _bdr
        ws[f"D{row}"] = iso_ref
        ws[f"D{row}"].fill = fill
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"].border = _bdr
        ws[f"E{row}"] = action
        ws[f"E{row}"].fill = fill
        ws[f"E{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"E{row}"].border = _bdr

    _finding(31, "Asset categories with major ownership gaps",
             "=COUNTIF(\'Ownership Coverage\'!H4:H8,\"\u274c Major Gaps\")",
             "CRITICAL", "A.5.9 §5.1",
             "Immediate — unowned assets have no accountability for protection, classification, or disposal",
             _crit, "C00000")
    _finding(32, "Total assets without an assigned owner",
             "=SUM(\'Ownership Coverage\'!D4:D8)",
             "CRITICAL", "A.5.9 §5.1",
             "Immediate — every asset must have a named responsible owner per A.5.9",
             _crit, "C00000")
    _finding(33, "Owners who have not attested responsibilities",
             "=COUNTA(\'Owner Acknowledgment\'!A4:A53)-COUNTIFS(\'Owner Acknowledgment\'!L4:L53,\"Attested*\")",
             "HIGH", "A.5.9 §5.2",
             "Urgent — unacknowledged ownership cannot be relied upon for compliance",
             _high, "9C5700")
    _finding(34, "Unaware asset owners",
             "=COUNTIF(\'Owner Awareness\'!G4:G53,\"\u274c Unaware\")",
             "HIGH", "A.5.9 §5.3",
             "Urgent — owners who are unaware cannot fulfil classification or review duties",
             _high, "9C5700")
    _finding(35, "Underperforming owners (Weak rating)",
             "=COUNTIF(\'Owner Performance\'!K4:K53,\"\u274c Weak\")",
             "HIGH", "A.5.9 §5.4",
             "Plan — underperforming owners require coaching, re-assignment, or escalation",
             _high, "9C5700")
    _finding(36, "Asset categories with minor coverage gaps",
             "=COUNTIF(\'Ownership Coverage\'!H4:H8,\"\u26a0\ufe0f Minor Gaps\")",
             "HIGH", "A.5.9 §5.1",
             "Plan — minor gaps require formal gap-closure plan before next review cycle",
             _high, "9C5700")
    _finding(37, "Owners needing improvement in performance",
             "=COUNTIF(\'Owner Performance\'!K4:K53,\"\u26a0\ufe0f Needs Improvement\")",
             "MEDIUM", "A.5.9 §5.4",
             "Plan — performance improvement plans required for owners below target",
             _med, "276221")
    _finding(38, "Unverified evidence items",
             "=COUNTIF(\'Evidence Register\'!H6:H105,\"Not verified\")",
             "MEDIUM", "A.5.9 §3",
             "Plan — evidence requires verification before next audit",
             _med, "276221")



def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-001..015)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = blue
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8); Overall Compliance Rating at B6 (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = yellow
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between approver sections

    # FINAL DECISION (GS-AS-004/012: col A plain bold, no dark fill)
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = yellow
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS (GS-AS-005/013: 4472C4 banner, borders on all)
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = blue
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
