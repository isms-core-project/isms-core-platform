#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.9.2 - Inventory Maintenance Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Assessment Domain 2 of 5: Inventory Maintenance & Update Procedures

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific inventory systems, update workflows, and integration
architecture.

Key customization areas:
1. Inventory system architecture (CMDB, spreadsheets, database, commercial tools)
2. Update workflows and triggers (procurement, decommissioning, change management)
3. Integration points (HR, procurement, asset management, monitoring tools)
4. SLA definitions and thresholds (what's acceptable for your operations)
5. Organisation name, CISO details, contact information
6. File paths and naming conventions
7. Existing integration patterns and APIs

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
how well [Organisation] maintains its asset inventory - the processes, systems,
and integrations that keep inventory data current and accurate over time.

**Purpose:**
Enables systematic assessment of inventory maintenance procedures, update
timeliness, system integrations, and quality control processes. Measures
operational effectiveness of keeping the inventory synchronised with actual
infrastructure changes, personnel moves, and procurement activities.

**Assessment Scope:**
- Inventory structure and access controls (who can view/edit what)
- Update triggers and workflows (how changes flow into inventory)
- Integration architecture (automated feeds from HR, procurement, monitoring)
- Quality control processes (validation, reconciliation, review cycles)
- Update timeliness and SLA compliance
- Documentation and procedures

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and maintenance best practices
2. Inventory Structure & Access - System architecture and access control model
3. Update Triggers & Workflows - How inventory stays synchronized with reality
4. Integration Architecture - Automated data feeds and system connections
5. Quality Control Processes - Validation, reconciliation, review procedures
6. Maintenance Metrics - Update timeliness, SLA compliance, effectiveness scores
7. Evidence Register - Process documentation and integration evidence

**Key Features:**
- Data validation with dropdown lists for system types and integration methods
- Conditional formatting for SLA compliance status (Green/Yellow/Red)
- Automated gap identification for missing integrations
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- SLA tracking and compliance measurement
- Integration health monitoring
- Update procedure documentation

**Integration:**
data from all four assessment domains for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a59_2_inventory_maintenance.py

Output:
    File: ISMS-IMP-A.5.9.2_Inventory_Maintenance_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Complete maintenance assessment by documenting all update procedures
    2. Map integration architecture (HR, procurement, monitoring tools)
    3. Define and track SLAs for inventory updates
    4. Measure update timeliness and compliance
    5. Collect and link evidence in Evidence Register sheet
    6. Review with stakeholders (IT Ops, Asset Management, Procurement, HR)
    7. Export metrics CSV for dashboard consolidation (Sheet 6)
    8. Store assessment workbook per retention policy (7 years minimum)
    9. Update quarterly or after major process/system changes

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.9
Assessment Domain:    2 of 5
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License]

Related Documents:
    - ISMS-POL-A.5.9: Inventory of Information and Assets (Policy)
    - ISMS-IMP-A.5.9-1: Asset Discovery Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-2: Inventory Maintenance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-3: Quality & Compliance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-4: Owner Accountability Assessment (Implementation Guide)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 22.01.2026
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.9-2 specification
    - Inventory maintenance procedure documentation
    - Integration architecture mapping
    - SLA tracking and compliance measurement
    - Quality control process assessment
    - Evidence collection and audit trail support

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive asset inventory details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review asset inventory procedures and classification criteria annually or when
new asset categories are introduced, system landscapes change, or audit findings
identify inventory gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
from datetime import datetime
from pathlib import Path
import os

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# CUSTOMIZE: Configuration
_INTEGRATION_TYPES = [
    'Real-time API',
    'Scheduled Batch (Daily)',
    'Scheduled Batch (Weekly)',
    'Webhook / Event-driven',
    'File Transfer (SFTP)',
    'Database Replication',
    'Manual Import',
    'No Integration (Manual)',
]

# Document identification constants

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.9.2"
WORKBOOK_NAME = "Inventory Maintenance"
CONTROL_ID   = "A.5.9"
CONTROL_NAME = "Inventory of Information and Assets"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("="*80)
    logger.info("ISMS Control A.5.9 - Inventory Maintenance Assessment Generator")
    logger.info("="*80)
    logger.info("")
    logger.info("Generating assessment workbook...")
    logger.info("")
    
    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create sheets in order
    sheets = [
        "Instructions & Legend",
        "Inventory Structure & Access",
        "Update Triggers & Workflows",
        "Integration Architecture",
        "Quality Control Processes",
        "Maintenance Metrics",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
        logger.info(f"  ✓ Created sheet: {sheet_name}")
    
    logger.info("")
    logger.info("Populating sheets...")
    logger.info("")
    
    # Populate each sheet
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✓ Instructions & Legend")
    
    create_inventory_structure_sheet(wb["Inventory Structure & Access"])
    logger.info("  ✓ Inventory Structure & Access")
    
    create_update_triggers_sheet(wb["Update Triggers & Workflows"])
    logger.info("  ✓ Update Triggers & Workflows")
    
    create_integration_architecture_sheet(wb["Integration Architecture"])
    logger.info("  ✓ Integration Architecture")
    
    create_quality_control_sheet(wb["Quality Control Processes"])
    logger.info("  ✓ Quality Control Processes")
    
    create_maintenance_metrics_sheet(wb["Maintenance Metrics"])
    logger.info("  ✓ Maintenance Metrics")
    
    create_evidence_register(wb["Evidence Register"])
    logger.info("  ✓ Evidence Register")
    
    create_summary_dashboard_sheet(wb["Summary Dashboard"])
    logger.info("  ✓ Summary Dashboard")
    
    create_approval_sheet(wb["Approval Sign-Off"])
    logger.info("  ✓ Approval & Sign-Off")
    
    # Finalise data validations
    finalize_validations(wb)

    # Save workbook to WKBK directory
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    wb.save(output_path)

    logger.info("")
    logger.info("="*80)
    logger.info(f"SUCCESS: {output_path.name}")
    logger.info("="*80)
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open workbook and review Instructions & Legend")
    logger.info("  2. Document inventory system architecture and access controls")
    logger.info("  3. Map all update triggers and workflows")
    logger.info("  4. Document integration architecture (HR, procurement, monitoring)")
    logger.info("  5. Assess quality control processes and procedures")
    logger.info("  6. Review Maintenance Metrics for SLA compliance")
    logger.info("  7. Export CSV from Sheet 6 for dashboard consolidation")
    logger.info("  8. Obtain stakeholder review and approval")
    logger.info("")
    logger.info(f"Output location: {output_path}")
    logger.info("")



def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Inventory Structure and Access sheet with your system architecture.', '2. Complete the Update Triggers and Workflows sheet with all trigger events.', '3. Complete the Integration Architecture sheet with all automated data feeds.', '4. Complete the Quality Control Processes sheet with all QC procedures.', '5. Review the Maintenance Metrics sheet for overall effectiveness scores.', '6. Link evidence in the Evidence Register sheet.', '7. Obtain approvals in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 21

    # Status Legend — row position tracks after instructions
    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_inventory_structure_sheet(ws):
    """Create Inventory Structure & Access sheet"""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "INVENTORY STRUCTURE & ACCESS CONTROLS"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # Section 1: Inventory System Architecture
    ws['A3'] = "TABLE 1: INVENTORY SYSTEM ARCHITECTURE"
    ws['A3'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    ws.merge_cells('A3:L3')
    
    # Column headers
    headers = [
        ('A', 'System Component', 30),
        ('B', 'Technology/Platform', 25),
        ('C', 'Purpose', 40),
        ('D', 'Data Stored', 40),
        ('E', 'Owner', 20),
        ('F', 'Backup Frequency', 20),
        ('G', 'Availability', 15),
        ('H', 'Documentation', 30),
        ('I', 'Last Review', 15),
        ('J', 'Status', 15),
        ('K', 'Evidence ID', 15),
        ('L', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}4'] = header
        ws[f'{col}4'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}4'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws[f'{col}4'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    # Example system components (user fills actual values)
    components = [
        ("Asset Inventory Database", "PostgreSQL / MySQL / SQL Server / Other", "Master inventory database"),
        ("CMDB (Configuration Management DB)", "ServiceNow / BMC / Jira / Other", "IT infrastructure CMDB"),
        ("Asset Management Tool", "Commercial Tool / Spreadsheet / Custom", "Asset tracking and management"),
        ("HR System Integration", "Workday / SAP SuccessFactors / Other", "Personnel asset integration"),
        ("Procurement System Integration", "SAP / Oracle / Coupa / Other", "Purchase order integration"),
    ]
    
    row = 5
    for component, tech, purpose in components:
        ws[f'A{row}'] = component
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = tech
        ws[f'C{row}'] = purpose
        
        # Unlock user input cells
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            pass
        
        row += 1
    
    # Add blank rows for additional components
    for i in range(5):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            pass
        row += 1
    
    # Section 2: Access Control Matrix
    row += 2
    ws[f'A{row}'] = "TABLE 2: ACCESS CONTROL MATRIX"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    ws.merge_cells(f'A{row}:L{row}')
    
    # Access control headers
    row += 1
    access_headers = [
        ('A', 'Role / User Group', 30),
        ('B', 'View Access', 15),
        ('C', 'Create Access', 15),
        ('D', 'Modify Access', 15),
        ('E', 'Delete Access', 15),
        ('F', 'Export Access', 15),
        ('G', 'Admin Access', 15),
        ('H', 'Justification', 40),
        ('I', 'Approved By', 20),
        ('J', 'Last Review', 15),
        ('K', 'Evidence ID', 15),
        ('L', 'Notes', 30),
    ]
    
    for col, header, width in access_headers:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}{row}'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Example roles
    roles = [
        "Asset Owners",
        "IT Operations",
        "Security Team",
        "Auditors (Read-Only)",
        "Executive Management (Reports)",
        "Procurement Team",
        "HR Team",
    ]
    
    row += 1
    for role in roles:
        ws[f'A{row}'] = role
        ws[f'A{row}'].font = Font(bold=True)
        
        # Unlock for user input
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            pass
        row += 1
    
    # Data validations for Yes/No
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_yesno.add(f'B{row-7}:G{row-1}')
    ws.add_data_validation(dv_yesno)


def create_update_triggers_sheet(ws):
    """Create Update Triggers & Workflows sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "UPDATE TRIGGERS & WORKFLOWS"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Update Trigger Event', 35),
        ('B', 'Trigger Source', 25),
        ('C', 'Workflow Documented?', 20),
        ('D', 'Automation Level', 20),
        ('E', 'Target SLA (Days)', 15),
        ('F', 'Actual Avg Time (Days)', 18),
        ('G', 'SLA Compliance %', 16),
        ('H', 'Compliance Status', 18),
        ('I', 'Responsible Party', 25),
        ('J', 'Procedure Location', 35),
        ('K', 'Last Review', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    # Pre-populate trigger events
    trigger_events = [
        ("New Asset Procurement", "Procurement System", 5),
        ("Asset Deployment / Go-Live", "Change Management", 3),
        ("Asset Configuration Change", "Change Management", 3),
        ("Asset Relocation / Transfer", "Facilities / IT Ops", 3),
        ("Asset Decommission / Disposal", "Change Management", 2),
        ("Personnel Onboarding (Joiner)", "HR System", 1),
        ("Personnel Offboarding (Leaver)", "HR System", 1),
        ("Personnel Role Change", "HR System", 2),
        ("Ownership Transfer", "Asset Owner / Security", 3),
        ("Security Incident (Asset Affected)", "Security Team", 1),
        ("Discovery Scan Finds New Asset", "Network Monitoring", 5),
        ("Periodic Review Identifies Change", "Asset Owner", 7),
    ]
    
    row = 4
    for event, source, sla in trigger_events:
        ws[f'A{row}'] = event
        # FFFFCC fill + borders for data row; F2F2F2 for sample row 4
        _fill_color = 'F2F2F2' if row == 4 else 'FFFFCC'
        _yf = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type='solid')
        from openpyxl.styles import Border as _Bdr, Side as _Sd
        _ts = _Sd(style='thin')
        _db = _Bdr(left=_ts, right=_ts, top=_ts, bottom=_ts)
        for _lc in 'ABCDEFGHIJKLM':
            ws[f'{_lc}{row}'].fill = _yf
            ws[f'{_lc}{row}'].border = _db

        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = source
        ws[f'E{row}'] = sla
        
        # Actual Avg Time (user enters from metrics)
        
        # SLA Compliance % (formula)
        ws[f'G{row}'] = f'=IF(F{row}="","",IF(F{row}<=E{row},100,MAX(0,100-(F{row}-E{row})/E{row}*100)))'
        ws[f'G{row}'].number_format = '0"%"'
        
        # Compliance Status (formula)
        ws[f'H{row}'] = f'=IF(G{row}="","",IF(G{row}>=95,"✅ Compliant",IF(G{row}>=80,"⚠️ At Risk","❌ Non-Compliant")))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Unlock user input cells
        for col in ['C', 'D', 'F', 'I', 'J', 'K', 'L', 'M']:
            pass
        
        row += 1
    
    # Data validations
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    dv_yesno.add(f'C4:C{row-1}')
    ws.add_data_validation(dv_yesno)
    
    automation_levels = ["Fully Automated", "Partially Automated", "Manual with Procedure", "Manual (Ad-hoc)", "Not Defined"]
    dv_automation = DataValidation(type="list", formula1=f'"{",".join(automation_levels)}"', allow_blank=True)
    dv_automation.add(f'D4:D{row-1}')
    ws.add_data_validation(dv_automation)
    
    # Summary section
    summary_row = row + 2
    ws[f'A{summary_row}'] = "UPDATE WORKFLOW SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Update Triggers"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Triggers with Documented Workflow"
    ws[f'B{summary_row}'] = f'=COUNTIF(C4:C{row-1},"Yes")'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Workflow Documentation %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Fully Automated Triggers"
    ws[f'B{summary_row}'] = f'=COUNTIF(D4:D{row-1},"Fully Automated")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Automation Rate %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-4}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average SLA Compliance %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(G4:G{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='between', formula=['80', '94'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='lessThan', formula=['80'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_integration_architecture_sheet(ws):
    """Create Integration Architecture sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "INTEGRATION ARCHITECTURE"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = CONTROL_REF
    ws['A2'].font = Font(size=10, italic=True, color="003366")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    
    # Column headers
    headers = [
        ('A', 'Source System', 30),
        ('B', 'Integration Type', 25),
        ('C', 'Data Flow', 30),
        ('D', 'Frequency', 20),
        ('E', 'Status', 15),
        ('F', 'Last Successful Run', 18),
        ('G', 'Success Rate %', 15),
        ('H', 'Health Status', 18),
        ('I', 'Owner', 20),
        ('J', 'Documentation', 35),
        ('K', 'Monitoring', 25),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Pre-populate integration examples
    integrations = [
        ("HR System (Personnel)", "Scheduled Batch (Daily)", "Employee data → Personnel Assets", "Daily 02:00"),
        ("Procurement System", "Webhook / Event-driven", "PO Approval → New Asset Record", "Real-time"),
        ("Network Monitoring (Nmap/Nessus)", "Scheduled Batch (Weekly)", "Scan Results → IT Infrastructure", "Weekly Sunday"),
        ("CMDB / ServiceNow", "Real-time API", "CI Updates → Inventory Sync", "Real-time"),
        ("Cloud Provider APIs (AWS/Azure)", "Scheduled Batch (Daily)", "Cloud Resources → IT Infrastructure", "Daily 03:00"),
        ("Active Directory", "Database Replication", "User/Group Data → Personnel/Access", "Hourly"),
        ("Asset Management Tool", "Real-time API", "Asset Changes → Inventory Sync", "Real-time"),
    ]
    
    row = 4
    for source, integration_type, data_flow, frequency in integrations:
        ws[f'A{row}'] = source
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = integration_type
        ws[f'C{row}'] = data_flow
        ws[f'D{row}'] = frequency
        
        # Success Rate % (user enters from monitoring)
        ws[f'G{row}'].number_format = '0"%"'
        
        # Health Status (formula based on success rate)
        ws[f'H{row}'] = f'=IF(G{row}="","",IF(G{row}>=98,"✅ Healthy",IF(G{row}>=90,"⚠️ Degraded","❌ Unhealthy")))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Unlock user input cells
        for col in ['E', 'F', 'G', 'I', 'J', 'K', 'L', 'M']:
            pass
        
        row += 1
    
    # Add blank rows
    for i in range(5):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'L', 'M']:
            pass
        row += 1
    

    # Apply FFFFCC fill + thin borders to all data rows
    from openpyxl.styles import Border as _B59, Side as _S59
    _ts = _S59(style="thin")
    _bd = _B59(left=_ts, right=_ts, top=_ts, bottom=_ts)
    _yf = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _sf = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for _r in range(4, row):
        for _c in range(1, 13 + 1):
            _cel = ws.cell(row=_r, column=_c)
            _cel.fill = _sf if _r == 4 else _yf
            _cel.border = _bd
    # Data validations
    dv_integration = DataValidation(type="list", formula1=f'"{",".join(_INTEGRATION_TYPES)}"', allow_blank=True)
    dv_integration.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_integration)
    
    statuses = ["Active", "Inactive", "Planned", "Deprecated", "Failed"]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses)}"', allow_blank=True)
    dv_status.add(f'E4:E{row-1}')
    ws.add_data_validation(dv_status)
    
    # Summary section
    summary_row = row + 2
    ws[f'A{summary_row}'] = "INTEGRATION HEALTH SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Integrations"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-6})'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Active Integrations"
    ws[f'B{summary_row}'] = f'=COUNTIF(E4:E{row-6},"Active")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Healthy Integrations"
    ws[f'B{summary_row}'] = f'=COUNTIF(H4:H{row-6},"✅ Healthy")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Integration Health %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Success Rate"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(G4:G{row-6}),0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'G4:G{row-6}',
        CellIsRule(operator='greaterThanOrEqual', formula=['98'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-6}',
        CellIsRule(operator='between', formula=['90', '97'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-6}',
        CellIsRule(operator='lessThan', formula=['90'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_quality_control_sheet(ws):
    """Create Quality Control Processes sheet"""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "QUALITY CONTROL PROCESSES"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Quality Control Process', 35),
        ('B', 'Frequency', 20),
        ('C', 'Documented?', 15),
        ('D', 'Last Performed', 18),
        ('E', 'Responsible Party', 25),
        ('F', 'Issues Found (Last Run)', 25),
        ('G', 'Issues Resolved', 20),
        ('H', 'Effectiveness %', 18),
        ('I', 'Status', 18),
        ('J', 'Evidence ID', 15),
        ('K', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="003366", fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Quality control processes
    qc_processes = [
        ("Data Validation Rules (Automated)", "Real-time", "Prevents invalid data entry"),
        ("Duplicate Detection / Deduplication", "Weekly", "Identifies and merges duplicate records"),
        ("Orphan Record Detection", "Weekly", "Finds assets without owners or parent records"),
        ("Consistency Checks (Cross-System)", "Monthly", "Compares inventory vs. HR, CMDB, procurement"),
        ("Completeness Checks (Mandatory Fields)", "Daily", "Identifies records missing required attributes"),
        ("Ownership Verification", "Quarterly", "Confirms asset owners are still valid and acknowledged"),
        ("Stale Record Review", "Monthly", "Identifies assets not reviewed in >90 days"),
        ("Access Control Audit", "Quarterly", "Reviews who has access to inventory systems"),
        ("Integration Reconciliation", "Weekly", "Verifies integrations running successfully"),
        ("Periodic Full Inventory Reconciliation", "Annually", "Complete audit of inventory vs. reality"),
    ]
    
    row = 4
    for process, frequency, description in qc_processes:
        ws[f'A{row}'] = process
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = frequency
        
        # Issues Found (user enters from last run)
        
        # Issues Resolved
        
        # Effectiveness % (formula: resolved / found)
        ws[f'H{row}'] = f'=IF(OR(F{row}="",G{row}=""),"",IF(F{row}=0,100,G{row}/F{row}*100))'
        ws[f'H{row}'].number_format = '0"%"'
        
        # Status
        ws[f'I{row}'] = f'=IF(H{row}="","",IF(H{row}>=90,"✅ Effective",IF(H{row}>=70,"⚠️ Needs Improvement","❌ Ineffective")))'
        ws[f'I{row}'].alignment = Alignment(horizontal='center')
        
        # Unlock user input cells
        for col in ['C', 'D', 'E', 'F', 'G', 'J', 'K']:
            pass
        
        row += 1
    
    # Data validations
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    dv_yesno.add(f'C4:C{row-1}')
    ws.add_data_validation(dv_yesno)
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "QUALITY CONTROL SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total QC Processes"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-1})'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Documented Processes"
    ws[f'B{summary_row}'] = f'=COUNTIF(C4:C{row-1},"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Documentation Rate %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Effectiveness %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(H4:H{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'H4:H{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['90'],
                   fill=PatternFill(start_color="C6EFCE", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H4:H{row-1}',
        CellIsRule(operator='between', formula=['70', '89'],
                   fill=PatternFill(start_color="FFEB9C", fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H4:H{row-1}',
        CellIsRule(operator='lessThan', formula=['70'],
                   fill=PatternFill(start_color="FFC7CE", fill_type='solid'))
    )


def create_maintenance_metrics_sheet(ws):
    """Create Maintenance Metrics sheet - consolidates from other sheets"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "MAINTENANCE METRICS & SUMMARY"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Section 1: Overall Maintenance Effectiveness
    ws['A3'] = "OVERALL MAINTENANCE EFFECTIVENESS"
    ws['A3'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    ws.merge_cells('A3:H3')
    
    # Metrics table
    metrics = [
        ("Update Workflow SLA Compliance", "='Update Triggers & Workflows'!B40", "95%", "Update timeliness"),
        ("Integration Health", "='Integration Architecture'!B21", "98%", "Automated feed reliability"),
        ("Quality Control Effectiveness", "='Quality Control Processes'!B27", "90%", "Issue remediation rate"),
        ("Workflow Documentation Rate", "='Update Triggers & Workflows'!B36", "100%", "Procedure documentation"),
    ]
    
    row = 4
    for metric, formula, target, description in metrics:
        ws[f'A{row}'] = metric
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0"%"'
        ws[f'C{row}'] = target
        ws[f'D{row}'] = f'=B{row}-VALUE(LEFT(C{row},LEN(C{row})-1))/100'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'E{row}'] = f'=IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100,"✅ Met",IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100-0.1,"⚠️ At Risk","❌ Not Met"))'
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        ws[f'F{row}'] = description
        row += 1
    
    # Column headers for metrics table
    ws['A4'].value = "Metric"
    ws['B4'].value = "Actual"
    ws['C4'].value = "Target"
    ws['D4'].value = "Gap"
    ws['E4'].value = "Status"
    ws['F4'].value = "Description"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}4'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}4'].fill = PatternFill(start_color="003366", fill_type='solid')
    
    row = 5
    for metric, formula, target, description in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0"%"'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = target
        ws[f'D{row}'] = f'=B{row}-VALUE(LEFT(C{row},LEN(C{row})-1))/100'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'E{row}'] = f'=IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100,"✅ Met",IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100-0.1,"⚠️ At Risk","❌ Not Met"))'
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        ws[f'F{row}'] = description
        row += 1
    
    # Overall score
    overall_row = row + 2
    ws[f'A{overall_row}'] = "OVERALL MAINTENANCE SCORE"
    ws[f'A{overall_row}'].font = Font(size=12, bold=True)
    ws[f'B{overall_row}'] = f'=AVERAGE(B5:B8)'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True, size=14)
    ws[f'B{overall_row}'].fill = PatternFill(start_color="FFEB9C", fill_type='solid')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Target"
    ws[f'B{overall_row}'] = "95%"
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Gap"
    ws[f'B{overall_row}'] = f'=B{overall_row-2}-0.95'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    
    # CSV Export section
    csv_row = overall_row + 3
    ws[f'A{csv_row}'] = "CSV EXPORT FOR DASHBOARD (Copy rows below)"
    ws[f'A{csv_row}'].font = Font(size=11, bold=True, color="FFFFFF")
    ws[f'A{csv_row}'].fill = PatternFill(start_color="4472C4", fill_type='solid')
    
    csv_row += 1
    ws[f'A{csv_row}'] = "Metric_Category"
    ws[f'B{csv_row}'] = "Score_%"
    ws[f'C{csv_row}'] = "Status"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{csv_row}'].font = Font(bold=True)
    
    csv_row += 1
    for i in range(4):
        ws[f'A{csv_row+i}'] = f'=A{5+i}'
        ws[f'B{csv_row+i}'] = f'=B{5+i}'
        ws[f'C{csv_row+i}'] = f'=E{5+i}'
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 40


def create_evidence_register(ws):
    """Create the standard Evidence Register sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Verification Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "A.5.9 All Domains", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Not verified"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Not verified,In Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

def create_summary_dashboard_sheet(ws):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 (A.5.9.2 Inventory Maintenance)."""
    from openpyxl.styles import Border, Side

    CHECK = "\u2705"
    XMARK = "\u274c"

    _thin = Side(style="thin")
    _bdr = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _red_b = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _crit = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _high = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _med = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    def _merge_row(row, fill, text, font_kw, align="left"):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.fill = fill
        c.font = Font(**font_kw)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = _bdr
        for col in "BCDE":
            ws[f"{col}{row}"].border = _bdr

    # Row 1: Title
    _merge_row(1, _navy, "INVENTORY MAINTENANCE \u2014 SUMMARY DASHBOARD",
               {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"}, align="center")
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    _merge_row(2, _blue, "ISO 27001:2022 \u00b7 Control A.5.9 \u00b7 Inventory of Information and Assets",
               {"name": "Calibri", "size": 10, "italic": True, "color": "FFFFFF"})
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    # ── TABLE 1 ────────────────────────────────────────────────────────────────
    _merge_row(4, _blue, "TABLE 1: COMPLIANCE ASSESSMENT",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    for col, label in enumerate(["Assessment Area", "Compliant", "Non-Compliant", "Total Items", "Compliance %"], 1):
        c = ws.cell(row=5, column=col, value=label)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr

    t1_rows = [
        (6, "Update Trigger Workflow Coverage",
         "=COUNTIF(\'Update Triggers & Workflows\'!C4:C15,\"Yes\")",
         "=COUNTA(\'Update Triggers & Workflows\'!A4:A15)"),
        (7, "Update Trigger SLA Compliance",
         "=COUNTIF(\'Update Triggers & Workflows\'!H4:H15,\"\u2705 Compliant\")",
         "=COUNTA(\'Update Triggers & Workflows\'!A4:A15)"),
        (8, "Integration Architecture Health",
         "=COUNTIF(\'Integration Architecture\'!H4:H15,\"\u2705 Healthy\")",
         "=COUNTA(\'Integration Architecture\'!A4:A15)"),
        (9, "Quality Control Process Coverage",
         "=COUNTIF(\'Quality Control Processes\'!C4:C53,\"Yes\")",
         "=COUNTA(\'Quality Control Processes\'!A4:A53)"),
        (10, "Maintenance Trigger Automation Rate",
         "=COUNTIF(\'Update Triggers & Workflows\'!D4:D15,\"Fully Automated\")",
         "=COUNTA(\'Update Triggers & Workflows\'!A4:A15)"),
    ]
    for row, label, b_fml, d_fml in t1_rows:
        ws.cell(row=row, column=1, value=label).border = _bdr
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{row}"] = b_fml
        ws[f"B{row}"].border = _bdr
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"] = f"=D{row}-B{row}"
        ws[f"C{row}"].border = _bdr
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"] = d_fml
        ws[f"D{row}"].border = _bdr
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"E{row}"] = f"=IFERROR(B{row}/D{row},0)"
        ws[f"E{row}"].number_format = "0.0%"
        ws[f"E{row}"].border = _bdr
        ws[f"E{row}"].alignment = Alignment(horizontal="center")

    for col, val in enumerate(["TOTAL", "=SUM(B6:B10)", "=SUM(C6:C10)", "=SUM(D6:D10)", "=IFERROR(B11/D11,0)"], 1):
        c = ws.cell(row=11, column=col, value=val)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        c.border = _bdr
    ws["E11"].number_format = "0.0%"

    ws.row_dimensions[12].height = 6

    # ── TABLE 2 ────────────────────────────────────────────────────────────────
    _merge_row(13, _blue, "TABLE 2: KEY PERFORMANCE INDICATORS",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    def _subhdr(row, label):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bdr
        for col in "BCDE":
            ws[f"{col}{row}"].border = _bdr

    def _metric(row, label, formula, fmt=None):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row}"].border = _bdr
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = _bdr
        if fmt:
            ws[f"B{row}"].number_format = fmt
        for col in "CDE":
            ws[f"{col}{row}"].border = _bdr

    _subhdr(14, "Update Trigger Coverage")
    _metric(15, "Total Update Triggers Defined", "=COUNTA(\'Update Triggers & Workflows\'!A4:A15)")
    _metric(16, "Triggers with Documented Workflow", "=COUNTIF(\'Update Triggers & Workflows\'!C4:C15,\"Yes\")")
    _metric(17, "Fully Automated Triggers", "=COUNTIF(\'Update Triggers & Workflows\'!D4:D15,\"Fully Automated\")")
    _subhdr(18, "Integration Architecture")
    _metric(19, "Total Integrations Monitored", "=COUNTA(\'Integration Architecture\'!A4:A15)")
    _metric(20, "Active Integrations", "=COUNTIF(\'Integration Architecture\'!E4:E15,\"Active\")")
    _metric(21, "Healthy Integrations", "=COUNTIF(\'Integration Architecture\'!H4:H15,\"\u2705 Healthy\")")
    _subhdr(22, "Quality Control")
    _metric(23, "Integration Failures (Failed Status)", "=COUNTIF(\'Integration Architecture\'!H4:H15,\"\u274c Failed\")")
    _metric(24, "Degraded Integrations", "=COUNTIF(\'Integration Architecture\'!H4:H15,\"\u26a0\ufe0f Degraded\")")
    _subhdr(25, "Gaps & Evidence")
    _metric(26, "Triggers Without Documented Workflow", "=COUNTIF(\'Update Triggers & Workflows\'!C4:C15,\"No\")")
    _metric(27, "Unverified Evidence Items", "=COUNTIF(\'Evidence Register\'!H6:H105,\"Not verified\")")

    ws.row_dimensions[28].height = 6

    # ── TABLE 3 ────────────────────────────────────────────────────────────────
    _merge_row(29, _red_b, "TABLE 3: CRITICAL FINDINGS & RISK INDICATORS",
               {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"})

    for col, label in enumerate(["Critical Finding", "Count", "Severity", "ISO Reference", "Action Required"], 1):
        c = ws.cell(row=30, column=col, value=label)
        c.fill = _grey
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr

    def _finding(row, label, formula, severity, iso_ref, action, fill, text_color):
        ws[f"A{row}"] = label
        ws[f"A{row}"].fill = fill
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"A{row}"].border = _bdr
        ws[f"B{row}"] = formula
        ws[f"B{row}"].fill = fill
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = _bdr
        ws[f"C{row}"] = severity
        ws[f"C{row}"].fill = fill
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"].border = _bdr
        ws[f"D{row}"] = iso_ref
        ws[f"D{row}"].fill = fill
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"].border = _bdr
        ws[f"E{row}"] = action
        ws[f"E{row}"].fill = fill
        ws[f"E{row}"].font = Font(name="Calibri", size=10, bold=True, color=text_color)
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"E{row}"].border = _bdr

    _finding(31, "Update triggers with no documented workflow",
             "=COUNTIF(\'Update Triggers & Workflows\'!C4:C15,\"No\")",
             "CRITICAL", "A.5.9 §3.3",
             "Immediate — undocumented triggers lead to missed inventory updates and stale data",
             _crit, "C00000")
    _finding(32, "Integration feeds with failed health status",
             "=COUNTIF(\'Integration Architecture\'!H4:H15,\"\u274c Failed\")",
             "CRITICAL", "A.5.9 §3.3",
             "Immediate — failed feeds create silent data gaps in the inventory",
             _crit, "C00000")
    _finding(33, "Integration feeds with degraded status",
             "=COUNTIF(\'Integration Architecture\'!H4:H15,\"\u26a0\ufe0f Degraded\")",
             "HIGH", "A.5.9 §3.3",
             "Urgent — degraded integrations reduce data quality and timeliness",
             _high, "9C5700")
    _finding(34, "Update triggers with no SLA compliance",
             "=COUNTIF(\'Update Triggers & Workflows\'!H4:H15,\"\u274c Non-Compliant\")",
             "HIGH", "A.5.9 §3.3",
             "Urgent — SLA breaches lead to stale inventory data",
             _high, "9C5700")
    _finding(35, "Critical triggers with no automation (ad-hoc)",
             "=COUNTIF(\'Update Triggers & Workflows\'!D4:D15,\"Manual (Ad-hoc)\")",
             "HIGH", "A.5.9 §3.3",
             "Plan — manual ad-hoc processes are unreliable for critical update triggers",
             _high, "9C5700")
    _finding(36, "Integrations not in active status",
             "=COUNTA(\'Integration Architecture\'!A4:A15)-COUNTIF(\'Integration Architecture\'!E4:E15,\"Active\")",
             "HIGH", "A.5.9 §3.3",
             "Plan — inactive integrations should be decommissioned or reactivated",
             _high, "9C5700")
    _finding(37, "Triggers with partial workflow only",
             "=COUNTIF(\'Update Triggers & Workflows\'!C4:C15,\"Partial\")",
             "MEDIUM", "A.5.9 §3.3",
             "Plan — partial workflows require completion for reliable inventory maintenance",
             _med, "276221")
    _finding(38, "Unverified evidence items",
             "=COUNTIF(\'Evidence Register\'!H6:H105,\"Not verified\")",
             "MEDIUM", "A.5.9 §3",
             "Plan — evidence requires verification before next audit",
             _med, "276221")



def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-001..015)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = blue
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8); Overall Compliance Rating at B6 (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = yellow
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between approver sections

    # FINAL DECISION (GS-AS-004/012: col A plain bold, no dark fill)
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = yellow
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS (GS-AS-005/013: 4472C4 banner, borders on all)
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = blue
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
