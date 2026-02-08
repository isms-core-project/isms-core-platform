#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.27.1 - Security Architecture Review Process Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.27: Secure System Architecture and Engineering
Assessment Domain 1 of 4: Security Architecture Review Process

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
the security architecture review process as defined in ISMS-POL-A.8.27.

**Purpose:**
Enables systematic assessment of [Organization]'s security architecture review
governance, process maturity, and effectiveness.

**Assessment Scope:**
- Review governance framework
- Review process and methodology
- Documentation templates
- SDLC integration
- Review effectiveness metrics
- Policy compliance

**Generated Workbook Structure:**
1. Instructions - Assessment methodology and guidance
2. Governance - Review governance framework assessment
3. Process - Review process maturity assessment
4. Templates - Documentation templates assessment
5. Integration - SDLC integration assessment
6. Metrics - Review effectiveness metrics
7. Compliance - Policy compliance scoring
8. GapRegister - Gap tracking and remediation
9. Dashboard - Summary view and status

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a827_1_architecture_review.py

Output:
    File: ISMS-IMP-A.8.27.1_Security_Architecture_Review_Process_YYYYMMDD.xlsx
    Location: ../90_workbooks/

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.27
Assessment Domain:    1 of 4
Framework Version:    1.0
Script Version:       1.0
Author:               ISMS Implementation Team
Date:                 [Date to be set]
Python Version:       3.8+

Related Documents:
    - ISMS-POL-A.8.27: Secure System Architecture and Engineering Principles
    - ISMS-IMP-A.8.27.1: Security Architecture Review Process (Specification)
    - NIST SP 800-160 Vol. 1 Rev. 1
    - INCOSE SE Handbook 5th Edition

--------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime
import os
import logging

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.27.1"
WORKBOOK_NAME = "Security Architecture Review Process"
CONTROL_ID = "A.8.27"
CONTROL_NAME = "Secure System Architecture and Engineering Principles"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# CONFIGURATION
# =============================================================================
CONFIG = {
    'organization': '[Organization]',
    'ciso_name': '[CISO Name]',
    'security_contact': '[security@organization.com]',

    'colors': {
        'header_bg': '1F4E79',       # Dark blue
        'header_text': 'FFFFFF',      # White
        'subheader_bg': '2E75B6',     # Medium blue
        'input_cell': 'E2EFDA',       # Light green (input)
        'formula_cell': 'FFF2CC',     # Light yellow (formula)
        'compliant': 'C6EFCE',        # Green
        'partial': 'FFEB9C',          # Yellow
        'non_compliant': 'FFC7CE',    # Red
        'high_risk': 'FF0000',        # Red
        'medium_risk': 'FFA500',      # Orange
        'low_risk': 'FFFFCC',         # Yellow
    },
}

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def get_styles():
    """Return dictionary of common styles"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'header': {
            'font': Font(bold=True, color=CONFIG['colors']['header_text'], size=12),
            'fill': PatternFill(start_color=CONFIG['colors']['header_bg'],
                               end_color=CONFIG['colors']['header_bg'], fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'subheader': {
            'font': Font(bold=True, color=CONFIG['colors']['header_text'], size=11),
            'fill': PatternFill(start_color=CONFIG['colors']['subheader_bg'],
                               end_color=CONFIG['colors']['subheader_bg'], fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'input': {
            'font': Font(size=11),
            'fill': PatternFill(start_color=CONFIG['colors']['input_cell'],
                               end_color=CONFIG['colors']['input_cell'], fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'formula': {
            'font': Font(size=11),
            'fill': PatternFill(start_color=CONFIG['colors']['formula_cell'],
                               end_color=CONFIG['colors']['formula_cell'], fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': thin_border,
        },
        'normal': {
            'font': Font(size=11),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
    }

def apply_style(cell, style_dict):
    """Apply style dictionary to cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create Instructions sheet"""
    styles = get_styles()

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ISMS-IMP-A.8.27.1 - Security Architecture Review Process Assessment"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    # Document info
    ws['A3'] = "Document Information"
    apply_style(ws['A3'], styles['subheader'])
    ws.merge_cells('A3:H3')

    info_data = [
        ('Document ID', DOCUMENT_ID),
        ('Control Reference', CONTROL_REF),
        ('Assessment Purpose', 'Evaluate security architecture review process maturity and compliance'),
        ('Generated Date', GENERATED_DATE),
        ('Organization', CONFIG['organization']),
    ]

    row = 4
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        apply_style(ws[f'A{row}'], styles['normal'])
        apply_style(ws[f'B{row}'], styles['input'])
        ws.merge_cells(f'B{row}:H{row}')
        row += 1

    row += 1
    ws[f'A{row}'] = "Assessment Instructions"
    apply_style(ws[f'A{row}'], styles['subheader'])
    ws.merge_cells(f'A{row}:H{row}')
    row += 1

    instructions = [
        "1. Review the Instructions sheet to understand the assessment methodology",
        "2. Complete the Governance sheet to assess review governance framework",
        "3. Complete the Process sheet to evaluate review process maturity",
        "4. Complete the Templates sheet to assess documentation templates",
        "5. Complete the Integration sheet to evaluate SDLC integration",
        "6. Complete the Metrics sheet to document effectiveness metrics",
        "7. Complete the Compliance sheet to score policy compliance",
        "8. Document all gaps in the GapRegister sheet",
        "9. Review the Dashboard for summary status",
    ]

    for instruction in instructions:
        ws[f'A{row}'] = instruction
        apply_style(ws[f'A{row}'], styles['normal'])
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

    row += 1
    ws[f'A{row}'] = "Rating Scale"
    apply_style(ws[f'A{row}'], styles['subheader'])
    ws.merge_cells(f'A{row}:H{row}')
    row += 1

    ratings = [
        ('1', 'Not performed or completely ineffective'),
        ('2', 'Performed ad-hoc, minimal effectiveness'),
        ('3', 'Defined process, moderate effectiveness'),
        ('4', 'Managed process, good effectiveness'),
        ('5', 'Optimised process, excellent effectiveness'),
    ]

    for rating, desc in ratings:
        ws[f'A{row}'] = rating
        ws[f'B{row}'] = desc
        apply_style(ws[f'A{row}'], styles['normal'])
        apply_style(ws[f'B{row}'], styles['normal'])
        ws.merge_cells(f'B{row}:H{row}')
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80


def create_governance_sheet(ws):
    """Create Governance assessment sheet"""
    styles = get_styles()

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Security Architecture Review - Governance Assessment"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['Gov-ID', 'Category', 'Requirement', 'Status', 'Evidence', 'Gap', 'Owner', 'Notes']
    widths = [10, 20, 40, 15, 40, 40, 20, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data validation for Status
    status_dv = DataValidation(
        type='list',
        formula1='"Implemented,Partial,Not Implemented,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)

    # Category validation
    category_dv = DataValidation(
        type='list',
        formula1='"Policy,Procedures,Roles,Authority,Exceptions"',
        allow_blank=True
    )
    ws.add_data_validation(category_dv)

    # Pre-populate governance requirements
    governance_reqs = [
        ('GOV-001', 'Policy', 'Architecture review policy documented'),
        ('GOV-002', 'Policy', 'Policy approved by CISO and Executive Management'),
        ('GOV-003', 'Procedures', 'Review procedures defined and maintained'),
        ('GOV-004', 'Procedures', 'Review methodology documented (STRIDE, etc.)'),
        ('GOV-005', 'Roles', 'Reviewer roles and responsibilities defined'),
        ('GOV-006', 'Roles', 'Approval authority documented'),
        ('GOV-007', 'Authority', 'CISO approval requirements specified'),
        ('GOV-008', 'Authority', 'Exception approval authority defined'),
        ('GOV-009', 'Exceptions', 'Exception process established'),
        ('GOV-010', 'Exceptions', 'Exception tracking and expiry defined'),
        ('GOV-011', 'Procedures', 'Review scope criteria defined'),
        ('GOV-012', 'Procedures', 'Mandatory review triggers documented'),
        ('GOV-013', 'Procedures', 'Documentation standards defined'),
        ('GOV-014', 'Procedures', 'Quality review process established'),
        ('GOV-015', 'Procedures', 'Continuous improvement process defined'),
    ]

    for row, (gov_id, category, requirement) in enumerate(governance_reqs, 4):
        ws.cell(row=row, column=1, value=gov_id)
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=requirement)

        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 3 else styles['normal'])

    # Apply validations
    status_dv.add(f'D4:D{3 + len(governance_reqs)}')
    category_dv.add(f'B4:B{3 + len(governance_reqs)}')

    # Freeze panes
    ws.freeze_panes = 'A4'


def create_process_sheet(ws):
    """Create Process assessment sheet"""
    styles = get_styles()

    ws.merge_cells('A1:H1')
    ws['A1'] = "Security Architecture Review - Process Assessment"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Proc-ID', 'Phase', 'Activity', 'Documented', 'Implemented', 'Evidence', 'Effectiveness', 'Notes']
    widths = [10, 15, 40, 12, 12, 30, 12, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    # Validations
    yes_no_dv = DataValidation(type='list', formula1='"Yes,Partial,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)

    rating_dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(rating_dv)

    phase_dv = DataValidation(
        type='list',
        formula1='"Trigger,Planning,Execution,Documentation,Approval,Follow-up"',
        allow_blank=True
    )
    ws.add_data_validation(phase_dv)

    # Pre-populate process activities
    process_activities = [
        ('PROC-001', 'Trigger', 'New system identification and intake'),
        ('PROC-002', 'Trigger', 'Major change request assessment'),
        ('PROC-003', 'Trigger', 'Architecture change impact evaluation'),
        ('PROC-004', 'Trigger', 'Third-party integration trigger assessment'),
        ('PROC-005', 'Trigger', 'Review threshold criteria application'),
        ('PROC-006', 'Planning', 'Review scope definition'),
        ('PROC-007', 'Planning', 'Reviewer assignment'),
        ('PROC-008', 'Planning', 'Timeline establishment'),
        ('PROC-009', 'Planning', 'Documentation gathering'),
        ('PROC-010', 'Execution', 'Architecture documentation review'),
        ('PROC-011', 'Execution', 'Threat modelling execution'),
        ('PROC-012', 'Execution', 'Security requirements validation'),
        ('PROC-013', 'Execution', 'Pattern compliance check'),
        ('PROC-014', 'Execution', 'Defence in depth assessment'),
        ('PROC-015', 'Execution', 'Risk assessment'),
        ('PROC-016', 'Documentation', 'Findings documentation'),
        ('PROC-017', 'Documentation', 'Risk rating assignment'),
        ('PROC-018', 'Documentation', 'Recommendations development'),
        ('PROC-019', 'Documentation', 'Review report completion'),
        ('PROC-020', 'Approval', 'Findings review with system owner'),
        ('PROC-021', 'Approval', 'Remediation plan agreement'),
        ('PROC-022', 'Approval', 'Approval decision and sign-off'),
        ('PROC-023', 'Follow-up', 'Finding tracking'),
        ('PROC-024', 'Follow-up', 'Remediation verification'),
        ('PROC-025', 'Follow-up', 'Lessons learned capture'),
    ]

    for row, (proc_id, phase, activity) in enumerate(process_activities, 4):
        ws.cell(row=row, column=1, value=proc_id)
        ws.cell(row=row, column=2, value=phase)
        ws.cell(row=row, column=3, value=activity)

        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 3 else styles['normal'])

    yes_no_dv.add(f'D4:E{3 + len(process_activities)}')
    rating_dv.add(f'G4:G{3 + len(process_activities)}')
    phase_dv.add(f'B4:B{3 + len(process_activities)}')

    ws.freeze_panes = 'A4'


def create_templates_sheet(ws):
    """Create Templates assessment sheet"""
    styles = get_styles()

    ws.merge_cells('A1:H1')
    ws['A1'] = "Security Architecture Review - Templates Assessment"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Temp-ID', 'Template', 'Version', 'Last Updated', 'Completeness', 'Usability', 'Gaps', 'Notes']
    widths = [10, 35, 10, 12, 12, 12, 35, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    rating_dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(rating_dv)

    templates = [
        ('TEMP-001', 'Security Architecture Document (SAD)'),
        ('TEMP-002', 'Threat Model Template'),
        ('TEMP-003', 'Architecture Review Checklist'),
        ('TEMP-004', 'Security Requirements Template'),
        ('TEMP-005', 'Risk Assessment Template'),
        ('TEMP-006', 'Architecture Decision Record (ADR)'),
        ('TEMP-007', 'Exception Request Form'),
        ('TEMP-008', 'Review Completion Report'),
    ]

    for row, (temp_id, template) in enumerate(templates, 4):
        ws.cell(row=row, column=1, value=temp_id)
        ws.cell(row=row, column=2, value=template)

        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 2 else styles['normal'])

    rating_dv.add(f'E4:F{3 + len(templates)}')

    ws.freeze_panes = 'A4'


def create_integration_sheet(ws):
    """Create Integration assessment sheet"""
    styles = get_styles()

    ws.merge_cells('A1:H1')
    ws['A1'] = "Security Architecture Review - SDLC Integration Assessment"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Int-ID', 'Integration', 'Trigger', 'Automated', 'Tracked', 'Enforced', 'Evidence', 'Notes']
    widths = [10, 30, 35, 12, 12, 12, 30, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    yes_no_dv = DataValidation(type='list', formula1='"Yes,Partial,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)

    integrations = [
        ('INT-001', 'Project initiation', 'New project intake form triggers review assessment'),
        ('INT-002', 'Architecture design phase', 'Design milestone triggers security review'),
        ('INT-003', 'Pre-development gate', 'Development cannot start without review approval'),
        ('INT-004', 'Pre-production release', 'Release blocked without architecture sign-off'),
        ('INT-005', 'Major change requests', 'Significant changes trigger re-review'),
        ('INT-006', 'Third-party integration', 'External service integration requires review'),
        ('INT-007', 'Cloud service adoption', 'New cloud services require architecture review'),
        ('INT-008', 'Post-incident review', 'Security incidents may trigger architecture reassessment'),
    ]

    for row, (int_id, integration, trigger) in enumerate(integrations, 4):
        ws.cell(row=row, column=1, value=int_id)
        ws.cell(row=row, column=2, value=integration)
        ws.cell(row=row, column=3, value=trigger)

        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 3 else styles['normal'])

    yes_no_dv.add(f'D4:F{3 + len(integrations)}')

    ws.freeze_panes = 'A4'


def create_metrics_sheet(ws):
    """Create Metrics assessment sheet"""
    styles = get_styles()

    ws.merge_cells('A1:H1')
    ws['A1'] = "Security Architecture Review - Effectiveness Metrics"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Met-ID', 'Metric', 'Period', 'Target', 'Actual', 'Trend', 'Action', 'Notes']
    widths = [10, 35, 15, 12, 12, 10, 35, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    trend_dv = DataValidation(type='list', formula1='"Up,Down,Stable"', allow_blank=True)
    ws.add_data_validation(trend_dv)

    metrics = [
        ('MET-001', 'Review coverage (% applicable projects reviewed)', '100%'),
        ('MET-002', 'Review timeliness (days from trigger to completion)', '<10 days'),
        ('MET-003', 'Finding closure rate (% high findings closed before release)', '100%'),
        ('MET-004', 'Bypass rate (% projects bypassing review)', '0%'),
        ('MET-005', 'Rework rate (% requiring re-review)', '<10%'),
        ('MET-006', 'Template compliance (% using approved templates)', '100%'),
        ('MET-007', 'Documentation completeness (average score)', '4.0/5'),
        ('MET-008', 'Stakeholder satisfaction (survey score)', '4.0/5'),
        ('MET-009', 'Time to approval (days from submission)', '<5 days'),
        ('MET-010', 'Finding recurrence (% findings repeating)', '<5%'),
    ]

    for row, (met_id, metric, target) in enumerate(metrics, 4):
        ws.cell(row=row, column=1, value=met_id)
        ws.cell(row=row, column=2, value=metric)
        ws.cell(row=row, column=4, value=target)

        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 2 else styles['normal'])

    trend_dv.add(f'F4:F{3 + len(metrics)}')

    ws.freeze_panes = 'A4'


def create_compliance_sheet(ws):
    """Create Compliance scoring sheet"""
    styles = get_styles()

    ws.merge_cells('A1:G1')
    ws['A1'] = "Security Architecture Review - Policy Compliance Scoring"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Comp-ID', 'Requirement', 'Source', 'Compliant', 'Evidence', 'Score', 'Notes']
    widths = [10, 40, 20, 12, 35, 10, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    compliant_dv = DataValidation(type='list', formula1='"Yes,Partial,No"', allow_blank=True)
    ws.add_data_validation(compliant_dv)

    compliance_reqs = [
        ('COMP-001', 'Security architecture review policy documented', 'POL-A.8.27 §2.2'),
        ('COMP-002', 'Review triggers defined for new systems', 'POL-A.8.27 §2.2.1'),
        ('COMP-003', 'Review triggers defined for major changes', 'POL-A.8.27 §2.2.1'),
        ('COMP-004', 'Threat modelling methodology adopted', 'POL-A.8.27 §2.2.1'),
        ('COMP-005', 'Security requirements validation required', 'POL-A.8.27 §2.2.1'),
        ('COMP-006', 'Pattern compliance review required', 'POL-A.8.27 §2.2.1'),
        ('COMP-007', 'Defence in depth assessment required', 'POL-A.8.27 §2.2.1'),
        ('COMP-008', 'Risk assessment and documentation required', 'POL-A.8.27 §2.2.1'),
        ('COMP-009', 'CISO/Security Architect approval required', 'POL-A.8.27 §2.2.1'),
        ('COMP-010', 'SAD documentation required', 'POL-A.8.27 §2.2.1'),
        ('COMP-011', 'Threat Model Report required', 'POL-A.8.27 §2.2.1'),
        ('COMP-012', 'Security Requirements Traceability Matrix', 'POL-A.8.27 §2.2.1'),
        ('COMP-013', 'Risk Assessment and Treatment Plan', 'POL-A.8.27 §2.2.1'),
        ('COMP-014', 'Architecture approval record', 'POL-A.8.27 §2.2.1'),
        ('COMP-015', 'Third-party architecture requirements', 'POL-A.8.27 §2.2.3'),
        ('COMP-016', 'Acquired systems security review', 'POL-A.8.27 §2.2.3'),
        ('COMP-017', 'Secure pattern catalogue maintained', 'POL-A.8.27 §2.2.2'),
        ('COMP-018', 'Pattern deviation requires approval', 'POL-A.8.27 §2.2.2'),
        ('COMP-019', 'Annual pattern review conducted', 'POL-A.8.27 §2.2.2'),
        ('COMP-020', 'Architecture review metrics tracked', 'POL-A.8.27 §4'),
    ]

    for row, (comp_id, requirement, source) in enumerate(compliance_reqs, 4):
        ws.cell(row=row, column=1, value=comp_id)
        ws.cell(row=row, column=2, value=requirement)
        ws.cell(row=row, column=3, value=source)

        # Score formula
        ws.cell(row=row, column=6, value=f'=IF(D{row}="Yes",100,IF(D{row}="Partial",50,0))')

        for col in range(1, 8):
            if col == 6:
                apply_style(ws.cell(row=row, column=col), styles['formula'])
            elif col > 3:
                apply_style(ws.cell(row=row, column=col), styles['input'])
            else:
                apply_style(ws.cell(row=row, column=col), styles['normal'])

    compliant_dv.add(f'D4:D{3 + len(compliance_reqs)}')

    # Overall score
    last_row = 3 + len(compliance_reqs) + 2
    ws.cell(row=last_row, column=5, value="Overall Compliance Score:")
    ws.cell(row=last_row, column=6, value=f'=AVERAGE(F4:F{3 + len(compliance_reqs)})')
    apply_style(ws.cell(row=last_row, column=5), styles['subheader'])
    apply_style(ws.cell(row=last_row, column=6), styles['formula'])

    ws.freeze_panes = 'A4'


def create_gap_register_sheet(ws):
    """Create Gap Register sheet"""
    styles = get_styles()

    ws.merge_cells('A1:J1')
    ws['A1'] = "Security Architecture Review - Gap Register"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Gap-ID', 'Source', 'Description', 'Risk', 'Remediation', 'Owner', 'Due Date', 'Status', 'Closure Date', 'Notes']
    widths = [10, 15, 40, 10, 40, 20, 12, 12, 12, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    risk_dv = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)

    status_dv = DataValidation(type='list', formula1='"Open,In Progress,Closed"', allow_blank=True)
    ws.add_data_validation(status_dv)

    source_dv = DataValidation(
        type='list',
        formula1='"Governance,Process,Templates,Integration,Metrics,Compliance"',
        allow_blank=True
    )
    ws.add_data_validation(source_dv)

    # Pre-populate 20 blank rows
    for row in range(4, 24):
        ws.cell(row=row, column=1, value=f'GAP-{row-3:03d}')
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 1 else styles['normal'])

    risk_dv.add('D4:D23')
    status_dv.add('H4:H23')
    source_dv.add('B4:B23')

    ws.freeze_panes = 'A4'


def create_dashboard_sheet(ws):
    """Create Dashboard summary sheet"""
    styles = get_styles()

    ws.merge_cells('A1:F1')
    ws['A1'] = "Security Architecture Review - Assessment Dashboard"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    # Summary section
    ws['A3'] = "Assessment Summary"
    apply_style(ws['A3'], styles['subheader'])
    ws.merge_cells('A3:F3')

    summary_items = [
        ('Assessment Date', GENERATED_DATE),
        ('Organization', CONFIG['organization']),
        ('Control Reference', CONTROL_REF),
        ('Overall Compliance Score', '=Compliance!F25'),  # Reference to compliance sheet
    ]

    row = 4
    for label, value in summary_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        apply_style(ws[f'A{row}'], styles['normal'])
        if '=' in str(value):
            apply_style(ws[f'B{row}'], styles['formula'])
        else:
            apply_style(ws[f'B{row}'], styles['input'])
        row += 1

    row += 1
    ws[f'A{row}'] = "Gap Summary"
    apply_style(ws[f'A{row}'], styles['subheader'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    gap_summary = [
        ('High Risk Gaps', '=COUNTIF(GapRegister!D:D,"High")'),
        ('Medium Risk Gaps', '=COUNTIF(GapRegister!D:D,"Medium")'),
        ('Low Risk Gaps', '=COUNTIF(GapRegister!D:D,"Low")'),
        ('Open Gaps', '=COUNTIF(GapRegister!H:H,"Open")'),
        ('In Progress Gaps', '=COUNTIF(GapRegister!H:H,"In Progress")'),
        ('Closed Gaps', '=COUNTIF(GapRegister!H:H,"Closed")'),
    ]

    for label, formula in gap_summary:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formula
        apply_style(ws[f'A{row}'], styles['normal'])
        apply_style(ws[f'B{row}'], styles['formula'])
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def main():
    """Main execution function"""
    logger.info("=" * 80)
    logger.info(f"ISMS Control {CONTROL_ID} - {WORKBOOK_NAME} Generator")
    logger.info("=" * 80)
    logger.info("")
    logger.info("Generating assessment workbook...")
    logger.info("")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create sheets
    sheets = [
        ("Instructions", create_instructions_sheet),
        ("Governance", create_governance_sheet),
        ("Process", create_process_sheet),
        ("Templates", create_templates_sheet),
        ("Integration", create_integration_sheet),
        ("Metrics", create_metrics_sheet),
        ("Compliance", create_compliance_sheet),
        ("GapRegister", create_gap_register_sheet),
        ("Dashboard", create_dashboard_sheet),
    ]

    for sheet_name, create_func in sheets:
        ws = wb.create_sheet(title=sheet_name)
        create_func(ws)
        logger.info(f"  ✓ Created sheet: {sheet_name}")

    logger.info("")

    # Determine output path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, '..', '90_workbooks')

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, OUTPUT_FILENAME)

    # Save workbook
    wb.save(output_path)

    logger.info(f"Workbook saved: {output_path}")
    logger.info("")
    logger.info("=" * 80)
    logger.info("Generation complete!")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()


# =============================================================================
# QA_VERIFIED: 2026-02-02
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation with standard structure
# =============================================================================
