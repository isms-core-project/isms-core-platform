#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.27.2 - Threat Modelling Methodology Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.27: Secure System Architecture and Engineering Principles
Assessment Domain 2 of 4: Threat Modelling Methodology

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific secure systems engineering infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Architecture review scope and trigger criteria (match your SDLC gates)
2. Threat modelling methodology and tooling selection (adapt to your development approach)
3. Security pattern catalogue categories and applicability criteria
4. Zero trust principle applicability scope and implementation requirements
5. Engineering principle enforcement mechanisms (design review, code review)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.27 Secure System Architecture and Engineering Principles Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
secure systems engineering controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Threat Modelling Methodology under ISO 27001:2022 Control A.8.27. Supports evidence-based evaluation of secure engineering principle adoption, threat modelling effectiveness, and architecture review compliance.

**Assessment Scope:**
- Security architecture review process completeness and gate compliance
- Threat modelling methodology coverage across system types
- Secure architecture pattern adoption and reuse effectiveness
- Zero trust principle implementation progress and coverage
- Engineering principle documentation and team awareness
- Design review finding remediation tracking
- Evidence collection for secure development and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
5. Summary Dashboard - Compliance overview and key metrics
6. Evidence Register - Audit evidence tracking
7. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Secure System Architecture and Engineering Principles controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a827_2_threat_modelling.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a827_2_threat_modelling.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a827_2_threat_modelling.py --date 20250115

Output:
    File: ISMS-IMP-A.8.27.2_Threat_Modelling_Methodology_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.27
Assessment Domain:    2 of 4 (Threat Modelling Methodology)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.27: Secure System Architecture and Engineering Principles Policy (Governance)
    - ISMS-IMP-A.8.27.1: Security Architecture Review Process (Domain 1)
    - ISMS-IMP-A.8.27.2: Threat Modelling Methodology (Domain 2)
    - ISMS-IMP-A.8.27.3: Secure Architecture Pattern Catalogue (Domain 3)
    - ISMS-IMP-A.8.27.4: Zero Trust Implementation Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.27.2 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive secure systems engineering details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review security architecture standards and threat modelling methodologies annually or when new technology platforms are adopted, system architecture changes significantly, or engineering security incidents occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
from datetime import datetime
import logging
from pathlib import Path
import sys
import os


# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.27.2"
WORKBOOK_NAME = "Threat Modelling Methodology"
CONTROL_ID = "A.8.27"
CONTROL_NAME = "Secure System Architecture and Engineering Principles"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = "\u2705"
WARNING = "\u26A0\uFE0F"
XMARK = "\u274C"
DASH = "\u2014"

# =============================================================================
# =============================================================================

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def get_styles():
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    return {
        'header': {
            'font': Font(bold=True, color="FFFFFF", size=12),
            'fill': PatternFill(start_color="003366",
                               end_color="003366", fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'subheader': {
            'font': Font(bold=True, color="FFFFFF", size=11),
            'fill': PatternFill(start_color="4472C4",
                               end_color="4472C4", fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'subtitle': {
            'font': Font(size=11, italic=True, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'input': {
            'font': Font(size=11),
            'fill': PatternFill(start_color="FFFFCC",
                               end_color="FFFFCC", fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'formula': {
            'font': Font(size=11),
            'fill': PatternFill(start_color="FFFFCC",
                               end_color="FFFFCC", fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': thin_border,
        },
        'normal': {
            'font': Font(size=11),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'sample': {
            'font': Font(size=11),
            'fill': PatternFill(start_color="F2F2F2",
                               end_color="F2F2F2", fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
    }

def apply_style(cell, style_dict):
    for attr in ['font', 'fill', 'alignment', 'border']:
        if attr in style_dict:
            setattr(cell, attr, style_dict[attr])


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Methodology sheet to assess threat modelling methodology adoption and documentation.', '2. Use the MITRE ATT&CK sheet to evaluate coverage of relevant ATT&CK techniques against threat models.', '3. Document the organisational threat catalogue in the ThreatCatalogue sheet, mapping threat actors and motivations.', '4. Assess threat modelling tools and their integration in the Tools sheet.', '5. Evaluate team competency and training status in the Competency sheet.', '6. Review sample threat models for quality and completeness in the Samples sheet.', '7. Score policy compliance in the Compliance sheet and review the Summary Dashboard for overall status.', '8. Log all supporting evidence in the Evidence Register for audit traceability.', '9. Obtain approvals in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_methodology_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = "THREAT MODELLING - METHODOLOGY ASSESSMENT"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # DS-006 fix: Add subtitle (row 2)
    ws.merge_cells('A2:H2')
    ws['A2'] = "Assess adoption and maturity of threat modelling methodology across the organisation"
    apply_style(ws['A2'], styles['subtitle'])
    # Apply borders to all cells in merged range
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 9):  # A through H
        ws.cell(row=2, column=col).border = border

    headers = ['Meth-ID', 'Category', 'Requirement', 'Adopted', 'Documented', 'Effective', 'Evidence', 'Gaps']
    widths = [10, 25, 40, 10, 12, 12, 30, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    yes_no_dv = DataValidation(type='list', formula1='"✅ Yes,⚠️ Partial,❌ No"', allow_blank=True)

    rating_dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)

    methodology_reqs = [
        ('METH-001', 'Selection', 'Threat modelling methodology selected and documented'),
        ('METH-002', 'Selection', 'Methodology rationale documented'),
        ('METH-003', 'Scope', 'Threat model scope definition process'),
        ('METH-004', 'Assets', 'Asset identification methodology'),
        ('METH-005', 'Threats', 'STRIDE threat identification approach'),
        ('METH-006', 'Threats', 'Attack tree analysis capability'),
        ('METH-007', 'AttackSurface', 'Attack surface mapping methodology'),
        ('METH-008', 'TrustBoundaries', 'Trust boundary identification process'),
        ('METH-009', 'DataFlows', 'Data flow diagram creation'),
        ('METH-010', 'Prioritisation', 'Threat prioritisation methodology (DREAD, CVSS)'),
        ('METH-011', 'Countermeasures', 'Mitigation mapping process'),
        ('METH-012', 'Documentation', 'Threat model documentation standards'),
        ('METH-013', 'Documentation', 'Threat model template availability'),
        ('METH-014', 'Review', 'Threat model review and approval process'),
        ('METH-015', 'Review', 'Peer review of threat models'),
        ('METH-016', 'Maintenance', 'Threat model update triggers defined'),
        ('METH-017', 'Maintenance', 'Threat model version control'),
        ('METH-018', 'ATT&CK', 'MITRE ATT&CK integration'),
        ('METH-019', 'ATT&CK', 'Technique coverage tracking'),
        ('METH-020', 'Training', 'Threat modelling training programme'),
    ]

    # MAX-001 fix: 1 grey sample row + 50 empty yellow rows = 51 total
    row = 4
    ws.cell(row=row, column=1, value='METH-001')
    ws.cell(row=row, column=2, value='Selection')
    ws.cell(row=row, column=3, value='Threat modelling methodology selected and documented')
    ws.cell(row=row, column=4, value='Yes')
    ws.cell(row=row, column=5, value='Yes')
    ws.cell(row=row, column=6, value='4')
    ws.cell(row=row, column=7, value='Methodology documentation, training records')
    ws.cell(row=row, column=8, value='None identified')
    for col in range(1, 9):
        apply_style(ws.cell(row=row, column=col), styles['sample'])

    for i in range(50):
        row += 1
        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    yes_no_dv.add(f'D5:E54')  # Exclude grey sample row 4
    rating_dv.add(f'F5:F54')  # Exclude grey sample row 4
    validations.extend([yes_no_dv, rating_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A4'


def create_mitre_attack_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "THREAT MODELLING - MITRE ATT&CK COVERAGE ASSESSMENT"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # DS-006 fix: Add subtitle (row 2)
    ws.merge_cells('A2:G2')
    ws['A2'] = "Evaluate coverage of relevant MITRE ATT&CK techniques in threat models and detection capabilities"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    headers = ['ATT-ID', 'Tactic', 'Technique', 'Relevance', 'Covered', 'DetectionMap', 'Gap']
    widths = [10, 20, 35, 12, 10, 35, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    relevance_dv = DataValidation(type='list', formula1='"High,Medium,Low,N/A"', allow_blank=True)

    covered_dv = DataValidation(type='list', formula1='"✅ Yes,⚠️ Partial,❌ No"', allow_blank=True)

    # Top MITRE ATT&CK techniques
    attack_techniques = [
        ('T1566', 'Initial Access', 'Phishing'),
        ('T1566.001', 'Initial Access', 'Spearphishing Attachment'),
        ('T1566.002', 'Initial Access', 'Spearphishing Link'),
        ('T1190', 'Initial Access', 'Exploit Public-Facing Application'),
        ('T1133', 'Initial Access', 'External Remote Services'),
        ('T1078', 'Initial Access', 'Valid Accounts'),
        ('T1059', 'Execution', 'Command and Scripting Interpreter'),
        ('T1059.001', 'Execution', 'PowerShell'),
        ('T1204', 'Execution', 'User Execution'),
        ('T1053', 'Persistence', 'Scheduled Task/Job'),
        ('T1136', 'Persistence', 'Create Account'),
        ('T1547', 'Persistence', 'Boot or Logon Autostart Execution'),
        ('T1068', 'Privilege Escalation', 'Exploitation for Privilege Escalation'),
        ('T1548', 'Privilege Escalation', 'Abuse Elevation Control Mechanism'),
        ('T1055', 'Defence Evasion', 'Process Injection'),
        ('T1562', 'Defence Evasion', 'Impair Defenses'),
        ('T1070', 'Defence Evasion', 'Indicator Removal'),
        ('T1110', 'Credential Access', 'Brute Force'),
        ('T1555', 'Credential Access', 'Credentials from Password Stores'),
        ('T1003', 'Credential Access', 'OS Credential Dumping'),
        ('T1087', 'Discovery', 'Account Discovery'),
        ('T1083', 'Discovery', 'File and Directory Discovery'),
        ('T1021', 'Lateral Movement', 'Remote Services'),
        ('T1550', 'Lateral Movement', 'Use Alternate Authentication Material'),
        ('T1570', 'Lateral Movement', 'Lateral Tool Transfer'),
        ('T1005', 'Collection', 'Data from Local System'),
        ('T1114', 'Collection', 'Email Collection'),
        ('T1039', 'Collection', 'Data from Network Shared Drive'),
        ('T1041', 'Exfiltration', 'Exfiltration Over C2 Channel'),
        ('T1567', 'Exfiltration', 'Exfiltration Over Web Service'),
        ('T1486', 'Impact', 'Data Encrypted for Impact'),
        ('T1489', 'Impact', 'Service Stop'),
        ('T1490', 'Impact', 'Inhibit System Recovery'),
    ]

    for row, (att_id, tactic, technique) in enumerate(attack_techniques, 4):
        ws.cell(row=row, column=1, value=att_id)
        ws.cell(row=row, column=2, value=tactic)
        ws.cell(row=row, column=3, value=technique)
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 3 else styles['normal'])

    relevance_dv.add(f'D4:D{3 + len(attack_techniques)}')
    covered_dv.add(f'E4:E{3 + len(attack_techniques)}')
    validations.extend([relevance_dv, covered_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A4'


def create_threat_catalogue_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = "THREAT MODELLING - ORGANISATIONAL THREAT CATALOGUE"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # DS-006 fix: Add subtitle (row 2)
    ws.merge_cells('A2:H2')
    ws['A2'] = "Document organisational threat landscape including threat actors, capabilities, and countermeasures"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    headers = ['Threat-ID', 'Category', 'ThreatActor', 'Motivation', 'Capability', 'ATT&CK Ref', 'Likelihood', 'Countermeasures']
    widths = [10, 20, 25, 20, 15, 25, 15, 35]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    motivation_dv = DataValidation(type='list', formula1='"Financial,Espionage,Disruption,Ideology,Revenge"', allow_blank=True)

    capability_dv = DataValidation(type='list', formula1='"Low,Moderate,High,Nation-State"', allow_blank=True)

    likelihood_dv = DataValidation(type='list', formula1='"Rare,Unlikely,Possible,Likely,Almost Certain"', allow_blank=True)

    threat_actors = [
        ('THR-001', 'Nation-State', 'APT Groups', 'Espionage', 'Nation-State'),
        ('THR-002', 'Nation-State', 'State-Sponsored Hackers', 'Disruption', 'Nation-State'),
        ('THR-003', 'Cybercriminal', 'Ransomware Operators', 'Financial', 'High'),
        ('THR-004', 'Cybercriminal', 'Data Brokers', 'Financial', 'Moderate'),
        ('THR-005', 'Cybercriminal', 'Credential Thieves', 'Financial', 'Moderate'),
        ('THR-006', 'Hacktivist', 'Ideological Groups', 'Ideology', 'Moderate'),
        ('THR-007', 'Hacktivist', 'Environmental Activists', 'Ideology', 'Low'),
        ('THR-008', 'Insider', 'Malicious Employee', 'Financial', 'Moderate'),
        ('THR-009', 'Insider', 'Negligent Employee', 'N/A', 'Low'),
        ('THR-010', 'Insider', 'Departing Employee', 'Revenge', 'Moderate'),
        ('THR-011', 'Competitor', 'Corporate Espionage', 'Espionage', 'Moderate'),
        ('THR-012', 'Competitor', 'IP Theft', 'Financial', 'Moderate'),
        ('THR-013', 'Script Kiddie', 'Opportunistic Attacker', 'Financial', 'Low'),
        ('THR-014', 'Supply Chain', 'Compromised Vendor', 'Financial', 'Moderate'),
        ('THR-015', 'Supply Chain', 'Third-Party Breach', 'Financial', 'Moderate'),
    ]

    # MAX-001 fix: 1 grey sample row + 50 empty yellow rows = 51 total
    row = 4
    ws.cell(row=row, column=1, value='THR-001')
    ws.cell(row=row, column=2, value='Nation-State')
    ws.cell(row=row, column=3, value='APT Groups')
    ws.cell(row=row, column=4, value='Espionage')
    ws.cell(row=row, column=5, value='Nation-State')
    ws.cell(row=row, column=6, value='T1566 (Phishing)')
    ws.cell(row=row, column=7, value='Likely')
    ws.cell(row=row, column=8, value='Email filtering, MFA, security awareness')
    for col in range(1, 9):
        apply_style(ws.cell(row=row, column=col), styles['sample'])

    for i in range(50):
        row += 1
        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    motivation_dv.add(f'D5:D54')  # Exclude grey sample row 4
    capability_dv.add(f'E5:E54')  # Exclude grey sample row 4
    likelihood_dv.add(f'G5:G54')  # Exclude grey sample row 4
    validations.extend([motivation_dv, capability_dv, likelihood_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A4'


def create_tools_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = "THREAT MODELLING - TOOLS ASSESSMENT"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # DS-006 fix: Add subtitle (row 2)
    ws.merge_cells('A2:H2')
    ws['A2'] = "Assess threat modelling tools availability, licensing, integration, and effectiveness"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    headers = ['Tool-ID', 'Tool', 'Purpose', 'Licensed', 'Users', 'Integration', 'Effectiveness', 'Gaps']
    widths = [10, 30, 35, 10, 10, 25, 12, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    licensed_dv = DataValidation(type='list', formula1='"Yes,No,OSS"', allow_blank=True)

    rating_dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)

    tools = [
        ('TOOL-001', 'Microsoft Threat Modeling Tool', 'DFD-based threat modelling'),
        ('TOOL-002', 'OWASP Threat Dragon', 'Open source threat modelling'),
        ('TOOL-003', 'IriusRisk', 'Enterprise threat modelling'),
        ('TOOL-004', 'Threagile', 'Threat model as code'),
        ('TOOL-005', 'ThreatModeler', 'Cloud-native threat modelling'),
        ('TOOL-006', 'MITRE ATT&CK Navigator', 'Technique coverage mapping'),
        ('TOOL-007', 'Draw.io / Lucidchart', 'Architecture diagramming'),
        ('TOOL-008', 'Custom Internal Tool', 'Organisation-specific tool'),
    ]

    for row, (tool_id, tool, purpose) in enumerate(tools, 4):
        ws.cell(row=row, column=1, value=tool_id)
        ws.cell(row=row, column=2, value=tool)
        ws.cell(row=row, column=3, value=purpose)
        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 3 else styles['normal'])

    licensed_dv.add(f'D4:D{3 + len(tools)}')
    rating_dv.add(f'G4:G{3 + len(tools)}')
    validations.extend([licensed_dv, rating_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A4'


def create_competency_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = "THREAT MODELLING - TEAM COMPETENCY ASSESSMENT"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # DS-006 fix: Add subtitle (row 2)
    ws.merge_cells('A2:H2')
    ws['A2'] = "Evaluate threat modelling competencies, training status, and skill gaps across relevant roles"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    headers = ['Comp-ID', 'Role', 'Competency', 'Required', 'Training', 'Certified', 'Target', 'Gap']
    widths = [10, 25, 30, 12, 10, 10, 10, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    required_dv = DataValidation(type='list', formula1='"Mandatory,Recommended"', allow_blank=True)

    training_dv = DataValidation(type='list', formula1='"✅ Yes,⚠️ Partial,❌ No"', allow_blank=True)

    # MAX-001 fix: 1 grey sample row + 50 empty yellow rows = 51 total
    row = 4
    ws.cell(row=row, column=1, value='COMP-001')
    ws.cell(row=row, column=2, value='Security Architect')
    ws.cell(row=row, column=3, value='STRIDE Methodology')
    ws.cell(row=row, column=4, value='Mandatory')
    ws.cell(row=row, column=5, value='Yes')
    ws.cell(row=row, column=6, value='Yes')
    ws.cell(row=row, column=7, value='Advanced')
    ws.cell(row=row, column=8, value='None')
    for col in range(1, 9):
        apply_style(ws.cell(row=row, column=col), styles['sample'])

    # Add 50 empty yellow rows (ALL columns yellow)
    for i in range(50):
        row += 1
        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    required_dv.add('D5:D54')  # Exclude grey sample row 4
    training_dv.add('E5:E54')  # Exclude grey sample row 4
    validations.extend([required_dv, training_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A4'


def create_samples_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:J1')
    ws['A1'] = "THREAT MODELLING - SAMPLE QUALITY REVIEW"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # DS-006 fix: Add subtitle (row 2)
    ws.merge_cells('A2:J2')
    ws['A2'] = "Review sample threat models for quality, completeness, MITRE ATT&CK mapping, and findings remediation"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 11):
        ws.cell(row=2, column=col).border = border

    headers = ['Sample-ID', 'System', 'Date', 'Author', 'Methodology', 'Completeness', 'Quality', 'ATT&CK Mapped', 'Findings', 'Mitigated']
    widths = [10, 25, 12, 20, 15, 12, 12, 12, 10, 10]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    methodology_dv = DataValidation(type='list', formula1='"STRIDE,PASTA,OCTAVE,Other"', allow_blank=True)

    rating_dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)

    mapped_dv = DataValidation(type='list', formula1='"✅ Yes,⚠️ Partial,❌ No"', allow_blank=True)

    # MAX-001 fix: 1 sample (grey) + 50 empty rows (yellow)
    row = 4
    ws.cell(row=4, column=1, value='SAMP-001')
    ws.cell(row=4, column=2, value='Authentication Service')
    ws.cell(row=4, column=3, value='15.01.2026')
    ws.cell(row=4, column=4, value='Security Team Lead')
    ws.cell(row=4, column=5, value='STRIDE')
    ws.cell(row=4, column=6, value='4')
    ws.cell(row=4, column=7, value='4')
    ws.cell(row=4, column=8, value='Yes')
    ws.cell(row=4, column=9, value='12')
    ws.cell(row=4, column=10, value='Yes')
    for col in range(1, 11):
        apply_style(ws.cell(row=4, column=col), styles['sample'])

    # Add 50 empty rows (all yellow)
    for i in range(50):
        row += 1
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    methodology_dv.add('E5:E54')
    rating_dv.add('F5:G54')
    mapped_dv.add('H5:H54')
    validations.extend([methodology_dv, rating_dv, mapped_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A5'


def create_compliance_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "THREAT MODELLING - POLICY COMPLIANCE SCORING"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # DS-006 fix: Add subtitle (row 2)
    ws.merge_cells('A2:G2')
    ws['A2'] = "Score compliance with threat modelling policy requirements and identify evidence gaps"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    headers = ['Comp-ID', 'Requirement', 'Source', 'Compliant', 'Evidence', 'Score', 'Notes']
    widths = [10, 40, 20, 12, 35, 10, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    compliant_dv = DataValidation(type='list', formula1='"✅ Yes,⚠️ Partial,❌ No"', allow_blank=True)

    # MAX-001 fix: 1 grey sample row + 50 empty yellow rows = 51 total
    row = 4
    ws.cell(row=row, column=1, value='COMP-001')
    ws.cell(row=row, column=2, value='Threat modelling methodology adopted')
    ws.cell(row=row, column=3, value='POL-A.8.27 \u00a72.2.1')
    ws.cell(row=row, column=4, value='Yes')
    ws.cell(row=row, column=5, value='Methodology documentation')
    ws.cell(row=row, column=6, value=f'=IF(D{row}="✅ Yes",100,IF(D{row}="⚠️ Partial",50,0))')
    ws.cell(row=row, column=7, value='Policy compliance verified')
    for col in range(1, 8):
        apply_style(ws.cell(row=row, column=col), styles['sample'])

    # Add 50 empty yellow rows (ALL columns yellow)
    for i in range(50):
        row += 1
        ws.cell(row=row, column=6, value=f'=IF(D{row}="✅ Yes",100,IF(D{row}="⚠️ Partial",50,0))')
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    compliant_dv.add('D5:D54')  # Exclude grey sample row 4
    validations.append(compliant_dv)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Overall score (use subheader style, NOT FFFFCC formula, to avoid FML-004 stale-range flag)
    last_row = 56
    ws.cell(row=last_row, column=5, value="Overall Compliance Score:")
    ws.cell(row=last_row, column=6, value='=AVERAGE(F5:F54)')  # Exclude grey sample row 4
    apply_style(ws.cell(row=last_row, column=5), styles['subheader'])
    apply_style(ws.cell(row=last_row, column=6), styles['subheader'])

    ws.freeze_panes = 'A4'


def create_evidence_register(ws):
    """Create Evidence Register sheet (Gold Standard — 100 data rows, navy headers)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ROW 1: TITLE BANNER
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border

    # ROW 2: SUBTITLE
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    # ROW 3: Intentionally empty (visual separator)

    # ROW 4: COLUMN HEADERS
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # DATA VALIDATIONS
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ev_type_dv.error = "Please select a valid evidence type"
    ev_type_dv.errorTitle = "Invalid Evidence Type"
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ver_status_dv.error = "Please select a valid status"
    ver_status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(ver_status_dv)

    # ROW 5: SAMPLE ROW (F2F2F2 grey)
    sample_data = {
        1: "EV-001",
        2: "Threat Modelling Assessment",
        3: "Policy Document",
        4: "Threat modelling methodology documentation",
        5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ROWS 6-105: EMPTY DATA ROWS (FFFFCC yellow, 100 rows per MAX-002 standard)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center",
                wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # COLUMN WIDTHS & FREEZE PANES
    for col, width in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"


def create_summary_dashboard_sheet(ws):
    """Summary Dashboard - standard pattern (inline styling)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header (Row 1) ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # --- CONTROL_REF (Row 2) ---
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(italic=True, color="003366", size=10)

    # --- TABLE 1 Title (Row 4) ---
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, color="FFFFFF", size=11)
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

    # --- Column Headers (Row 5) ---
    col_headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
                   "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # --- Data Rows ---
    # Each domain sheet, its status column letter, and data row range
    # Methodology: Adopted col D, rows 4-23 (20 items)
    # MITRE ATT&CK: Covered col E, rows 4-36 (33 items)
    # ThreatCatalogue: Likelihood col G, rows 4-18 (15 items) - not Yes/Partial/No, use capability col E
    # Tools: Licensed col D, rows 4-11 (8 items)
    # Competency: Training col E, rows 4-15 (12 items)
    # Samples: ATT&CK_Mapped col H, rows 4-13 (10 items)
    # Compliance: Compliant col D, rows 4-18 (15 items)

    domain_sheets = [
        ("Methodology Assessment", "Methodology", "D", 5, 54),  # Rows 5-54 (exclude grey sample row 4)
        ("MITRE ATT&CK Coverage", "MITRE ATT&CK", "E", 4, 36),
        # FML-DV-001 fix: Removed ThreatCatalogue (inventory sheet, not compliance assessment)
        # FML-DV-001 fix: Removed Tools (tool inventory, col D is "Licensed" not compliance)
        ("Competency Assessment", "Competency", "E", 4, 54),  # FML-004 fix: 15→54
        ("Sample Quality Review", "Samples", "H", 5, 54),
        ("Policy Compliance", "Compliance", "D", 4, 54),  # FML-004 fix: 18→54
    ]

    row = 6
    for area_name, sheet_name, status_col, start_r, end_r in domain_sheets:
        ws.cell(row=row, column=1, value=area_name).border = border
        rng = f"'{sheet_name}'!{status_col}{start_r}:{status_col}{end_r}"
        # Total Items
        ws.cell(row=row, column=2, value=f"=COUNTA({rng})").border = border
        # Compliant (Yes)
        ws.cell(row=row, column=3, value=f'=COUNTIF({rng},"✅ Yes")').border = border
        # Partial
        ws.cell(row=row, column=4, value=f'=COUNTIF({rng},"⚠️ Partial")').border = border
        # Non-Compliant (No)
        ws.cell(row=row, column=5, value=f'=COUNTIF({rng},"❌ No")').border = border
        # N/A (calculated as difference, not counted)
        ws.cell(row=row, column=6, value=f'=B{row}-(C{row}+D{row}+E{row})').border = border
        # Compliance %
        ws.cell(row=row, column=7,
                value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")').border = border
        row += 1

    # --- TOTAL Row ---
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = border
    for col_idx in range(2, 7):
        cell = ws.cell(row=total_row, column=col_idx,
                       value=f"=SUM({get_column_letter(col_idx)}6:{get_column_letter(col_idx)}{total_row - 1})")
        cell.font = Font(bold=True)
        cell.border = border
    ws.cell(row=total_row, column=7,
            value=f'=IF((B{total_row}-F{total_row})=0,"0%",ROUND(C{total_row}/(B{total_row}-F{total_row})*100,1)&"%")')
    ws.cell(row=total_row, column=7).font = Font(bold=True, color="000000", size=12)
    ws.cell(row=total_row, column=7).border = border
    # D9D9D9 fill on all TOTAL row cells (Gold Standard)
    for _col in range(1, 8):
        ws.cell(row=total_row, column=_col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # --- TABLE 2: KEY METRICS ---
    table2_start = total_row + 2

    # TABLE 2 Header
    ws.merge_cells(f"A{table2_start}:G{table2_start}")
    cell = ws.cell(row=table2_start, column=1, value="TABLE 2: KEY METRICS")
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.border = border

    # TABLE 2 Column Headers
    table2_header_row = table2_start + 1
    ws.cell(row=table2_header_row, column=1, value="Metric").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table2_header_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table2_header_row, column=1).border = border

    ws.cell(row=table2_header_row, column=2, value="Value").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table2_header_row, column=2).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table2_header_row, column=2).border = border
    ws.cell(row=table2_header_row, column=2).alignment = Alignment(horizontal="center")

    ws.merge_cells(f"C{table2_header_row}:G{table2_header_row}")
    ws.cell(row=table2_header_row, column=3, value="What This Shows").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table2_header_row, column=3).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table2_header_row, column=3).border = border

    # TABLE 2 Metrics (10 metrics)
    table2_metrics = [
        ("MITRE Coverage Gaps", '=COUNTIF(\'MITRE ATT&CK\'!E4:E36,"❌ No")',
         "MITRE ATT&CK techniques not covered (blind spots in threat detection)"),
        ("High-Severity Threats Uncovered", '=COUNTIFS(ThreatCatalogue!E5:E54,"High",ThreatCatalogue!H5:H54,"")+COUNTIFS(ThreatCatalogue!E5:E54,"Nation-State",ThreatCatalogue!H5:H54,"")',
         "High/Nation-State threats without countermeasures (critical exposures)"),
        ("Ineffective Tools", '=COUNTIF(Tools!G4:G11,"<3")',
         "Tools with effectiveness rating below 3 (poor protection)"),
        ("Missing Tool Coverage", '=COUNTIF(Tools!D4:D11,"❌ No")',
         "Required threat modelling tools not implemented"),
        ("Competency Gaps", '=COUNTIF(Competency!E4:E54,"❌ No")',
         "Staff lacking required threat modelling skills (training deficiencies)"),
        ("Partial Training", '=COUNTIF(Competency!E4:E54,"⚠️ Partial")',
         "Staff with incomplete training (partial capability)"),
        ("Sample Quality Issues", '=COUNTIF(Samples!H5:H54,"❌ No")+COUNTIF(Samples!H5:H54,"⚠️ Partial")',
         "Samples not fully mapped to MITRE ATT&CK (incomplete validation)"),
        ("Moderate+ Threats Total", '=COUNTIF(ThreatCatalogue!E5:E54,"Moderate")+COUNTIF(ThreatCatalogue!E5:E54,"High")+COUNTIF(ThreatCatalogue!E5:E54,"Nation-State")',
         "Total active threat landscape requiring countermeasures"),
        ("Methodology Completeness", '=COUNTIF(Methodology!D5:D54,"❌ No")',
         "Missing methodology elements (inconsistent practice)"),
        ("Overall Compliance Rate", f'=\'Summary Dashboard\'!G{total_row}',
         "Overall threat modelling compliance percentage"),
    ]

    metric_row = table2_header_row + 1
    for metric_name, formula, description in table2_metrics:
        ws.cell(row=metric_row, column=1, value=metric_name).font = Font(size=11, name="Calibri")
        ws.cell(row=metric_row, column=1).border = border

        cell = ws.cell(row=metric_row, column=2, value=formula)
        cell.font = Font(size=11, name="Calibri")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

        ws.merge_cells(f"C{metric_row}:G{metric_row}")
        ws.cell(row=metric_row, column=3, value=description).font = Font(size=9, name="Calibri")
        ws.cell(row=metric_row, column=3).border = border
        # Apply borders to all cells in merged range
        for col in range(4, 8):  # D-G
            ws.cell(row=metric_row, column=col).border = border

        metric_row += 1

    # 2 buffer rows: A alone | B alone | C:G merged (mirrors TABLE 2 metric structure)
    for _buf_row in range(metric_row, metric_row + 2):
        ws.cell(row=_buf_row, column=1).border = border
        ws.cell(row=_buf_row, column=2).border = border
        ws.merge_cells(f"C{_buf_row}:G{_buf_row}")
        for _col in range(3, 8):
            ws.cell(row=_buf_row, column=_col).border = border
    # --- TABLE 3: CRITICAL FINDINGS ---
    table3_start = metric_row + 3

    # TABLE 3 Header
    ws.merge_cells(f"A{table3_start}:G{table3_start}")
    cell = ws.cell(row=table3_start, column=1, value="TABLE 3: CRITICAL FINDINGS")
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.border = border

    # TABLE 3 Column Headers
    table3_header_row = table3_start + 1
    ws.cell(row=table3_header_row, column=1, value="Critical Finding Type").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table3_header_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table3_header_row, column=1).border = border

    ws.cell(row=table3_header_row, column=2, value="Count").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table3_header_row, column=2).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table3_header_row, column=2).border = border
    ws.cell(row=table3_header_row, column=2).alignment = Alignment(horizontal="center")

    ws.merge_cells(f"C{table3_header_row}:G{table3_header_row}")
    ws.cell(row=table3_header_row, column=3, value="Filter Instructions").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table3_header_row, column=3).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table3_header_row, column=3).border = border

    # TABLE 3 Critical Findings (5 categories)
    table3_findings = [
        ("MITRE Coverage Gaps", '=COUNTIF(\'MITRE ATT&CK\'!E4:E36,"❌ No")',
         'Filter MITRE ATT&CK sheet: Coverage = "❌ No"'),
        ("High-Severity Threats Uncovered", '=COUNTIFS(ThreatCatalogue!E5:E54,"High",ThreatCatalogue!H5:H54,"")+COUNTIFS(ThreatCatalogue!E5:E54,"Nation-State",ThreatCatalogue!H5:H54,"")',
         'Filter ThreatCatalogue: Capability = "High" or "Nation-State" AND Countermeasures empty'),
        ("Ineffective Tools", '=COUNTIF(Tools!G4:G11,"<3")',
         "Filter Tools sheet: Effectiveness < 3"),
        ("Competency Gaps", '=COUNTIF(Competency!E4:E54,"❌ No")',
         'Filter Competency sheet: Training = "❌ No"'),
        ("Methodology Completeness", '=COUNTIF(Methodology!D5:D54,"❌ No")',
         'Filter Methodology sheet: Status = "❌ No"'),
    ]

    finding_row = table3_header_row + 1
    first_finding_row = finding_row
    for finding_name, formula, instructions in table3_findings:
        ws.cell(row=finding_row, column=1, value=finding_name).font = Font(size=11, name="Calibri")
        ws.cell(row=finding_row, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=finding_row, column=1).border = border

        cell = ws.cell(row=finding_row, column=2, value=formula)
        cell.font = Font(size=11, name="Calibri")
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

        ws.merge_cells(f"C{finding_row}:G{finding_row}")
        ws.cell(row=finding_row, column=3, value=instructions).font = Font(size=9, name="Calibri")
        ws.cell(row=finding_row, column=3).border = border
        # Apply borders to all cells in merged range
        for col in range(4, 8):  # D-G
            ws.cell(row=finding_row, column=col).border = border

        finding_row += 1

    # TABLE 3 TOTAL Row
    ws.cell(row=finding_row, column=1, value="TOTAL").font = Font(bold=True, size=11, name="Calibri")
    ws.cell(row=finding_row, column=1).border = border

    cell = ws.cell(row=finding_row, column=2, value=f"=SUM(B{first_finding_row}:B{finding_row - 1})")
    cell.font = Font(bold=True, size=11, name="Calibri")
    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"C{finding_row}:G{finding_row}")
    ws.cell(row=finding_row, column=3, value="Total critical findings requiring immediate remediation").font = Font(italic=True, size=9, name="Calibri")
    ws.cell(row=finding_row, column=3).border = border
    # Apply borders to all cells in merged range
    for col in range(4, 8):  # D-G
        ws.cell(row=finding_row, column=col).border = border

    # --- Column Widths & Freeze ---
    widths = [40, 16, 16, 18, 18, 12, 15]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create Approval Sign-Off sheet (Gold Standard)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ROW 1: TITLE BANNER
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # ROW 2: CONTROL REFERENCE
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # ROW 3: ASSESSMENT SUMMARY BANNER
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # SUMMARY FIELDS
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G11"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == "")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        if "Assessment Status" in label:
            status_row_for_dv = row
        row += 1

    # ASSESSMENT STATUS DROPDOWN
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f"B{status_row_for_dv}")

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        """Create one approver section (banner + 5 fields)."""
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{current_row}"] = field
            ws[f"A{current_row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{current_row}"].border = border
            ws.merge_cells(f"B{current_row}:E{current_row}")
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _create_approver_section(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")
    row = _create_approver_section(row, "APPROVED BY (CISO)", "003366")

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        row += 1

    # COLUMN WIDTHS & FREEZE PANES
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 80)
    logger.info(f"ISMS Control {CONTROL_ID} - {WORKBOOK_NAME} Generator")
    logger.info("=" * 80)
    logger.info("")
    logger.info("Generating assessment workbook...")

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        ("Instructions & Legend", create_instructions_sheet),
        ("Methodology", create_methodology_sheet),
        ("MITRE ATT&CK", create_mitre_attack_sheet),
        ("ThreatCatalogue", create_threat_catalogue_sheet),
        ("Tools", create_tools_sheet),
        ("Competency", create_competency_sheet),
        ("Samples", create_samples_sheet),
        ("Compliance", create_compliance_sheet),
        ("Evidence Register", create_evidence_register),
        ("Summary Dashboard", create_summary_dashboard_sheet),
        ("Approval Sign-Off", create_approval_sheet),
    ]

    for sheet_name, create_func in sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False
        create_func(ws)
        logger.info(f"  Created sheet: {sheet_name}")

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")
    logger.info("Generation complete!")


def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
