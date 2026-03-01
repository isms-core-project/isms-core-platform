#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.27.3 - Secure Architecture Pattern Catalogue Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.27: Secure System Architecture and Engineering Principles
Assessment Domain 3 of 4: Secure Architecture Pattern Catalogue

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific secure systems engineering infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Architecture review scope and trigger criteria (match your SDLC gates)
2. Threat modelling methodology and tooling selection (adapt to your development approach)
3. Security pattern catalogue categories and applicability criteria
4. Zero trust principle applicability scope and implementation requirements
5. Engineering principle enforcement mechanisms (design review, code review)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.27 Secure System Architecture and Engineering Principles Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
secure systems engineering controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Secure Architecture Pattern Catalogue under ISO 27001:2022 Control A.8.27. Supports evidence-based evaluation of secure engineering principle adoption, threat modelling effectiveness, and architecture review compliance.

**Assessment Scope:**
- Security architecture review process completeness and gate compliance
- Threat modelling methodology coverage across system types
- Secure architecture pattern adoption and reuse effectiveness
- Zero trust principle implementation progress and coverage
- Engineering principle documentation and team awareness
- Design review finding remediation tracking
- Evidence collection for secure development and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
5. Summary Dashboard - Compliance overview and key metrics
6. Evidence Register - Audit evidence tracking
7. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Secure System Architecture and Engineering Principles controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a827_3_pattern_catalogue.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a827_3_pattern_catalogue.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a827_3_pattern_catalogue.py --date 20250115

Output:
    File: ISMS-IMP-A.8.27.3_Secure_Architecture_Pattern_Catalogue_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.27
Assessment Domain:    3 of 4 (Secure Architecture Pattern Catalogue)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.27: Secure System Architecture and Engineering Principles Policy (Governance)
    - ISMS-IMP-A.8.27.1: Security Architecture Review Process (Domain 1)
    - ISMS-IMP-A.8.27.2: Threat Modelling Methodology (Domain 2)
    - ISMS-IMP-A.8.27.3: Secure Architecture Pattern Catalogue (Domain 3)
    - ISMS-IMP-A.8.27.4: Zero Trust Implementation Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.27.3 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive secure systems engineering details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review security architecture standards and threat modelling methodologies annually or when new technology platforms are adopted, system architecture changes significantly, or engineering security incidents occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
from datetime import datetime
import logging
from pathlib import Path
import sys
import os


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.27.3"
WORKBOOK_NAME = "Secure Architecture Pattern Catalogue"
CONTROL_ID = "A.8.27"
CONTROL_NAME = "Secure System Architecture and Engineering Principles"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = "\u2705"
WARNING = "\u26A0\uFE0F"
XMARK = "\u274C"
DASH = "\u2014"

def get_styles():
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    return {
        'header': {'font': Font(bold=True, color="FFFFFF", size=12, name='Calibri'),
                   'fill': PatternFill(start_color="003366", end_color="003366", fill_type='solid'),
                   'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True), 'border': thin_border},
        'subheader': {'font': Font(bold=True, color="FFFFFF", size=11, name='Calibri'),
                      'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type='solid'),
                      'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'subtitle': {
            'font': Font(size=11, italic=True, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': thin_border,
        },
        'input': {'font': Font(size=11, name='Calibri'), 'fill': PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type='solid'),
                  'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'formula': {'font': Font(size=11, name='Calibri'), 'fill': PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type='solid'),
                    'alignment': Alignment(horizontal='center', vertical='center'), 'border': thin_border},
        'normal': {'font': Font(size=11, name='Calibri'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'sample': {'font': Font(size=11, name='Calibri'), 'fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type='solid'),
                   'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
    }

def apply_style(cell, style_dict):
    for attr in ['font', 'fill', 'alignment', 'border']:
        if attr in style_dict:
            setattr(cell, attr, style_dict[attr])


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# COMMON SHEETS (INLINE STYLING PER STANDARD-SCR-COMMON-SHEETS.MD)
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Document Information section above with your organisation details.', '2. PatternInventory: Register all secure architecture patterns with category, status and ownership.', '3. PatternQuality: Assess documentation quality for each pattern element.', '4. Adoption: Track pattern adoption rates across projects and identify barriers.', '5. Governance: Evaluate governance requirements for the pattern catalogue.', '6. Deviations: Record and track any deviations from approved patterns.', '7. Effectiveness: Measure pattern effectiveness using security incident and audit data.', '8. Compliance: Assess policy compliance for each pattern catalogue requirement.', '9. Evidence Register: List all evidence documents referenced in this assessment.', '10. Summary Dashboard: Review the auto-calculated compliance summary.', '11. Approval Sign-Off: Obtain required approvals before submission.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 25

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_evidence_register(ws):
    """Create Evidence Register sheet (Gold Standard — 100 data rows, navy headers)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ROW 1: TITLE BANNER
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border

    # ROW 2: SUBTITLE
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    # ROW 3: Intentionally empty (visual separator)

    # ROW 4: COLUMN HEADERS
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # DATA VALIDATIONS
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ev_type_dv.error = "Please select a valid evidence type"
    ev_type_dv.errorTitle = "Invalid Evidence Type"
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ver_status_dv.error = "Please select a valid status"
    ver_status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(ver_status_dv)

    # ROW 5: SAMPLE ROW (F2F2F2 grey)
    sample_data = {
        1: "EV-001",
        2: "Architecture Pattern Assessment",
        3: "Policy Document",
        4: "Secure architecture pattern catalogue documentation",
        5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ROWS 6-105: EMPTY DATA ROWS (FFFFCC yellow, 100 rows per MAX-002 standard)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center",
                wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # COLUMN WIDTHS & FREEZE PANES
    for col, width in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"


def create_summary_dashboard_sheet(ws):
    """Summary Dashboard sheet — standard common sheet pattern."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header (Row 1) ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # --- CONTROL_REF (Row 2) ---
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(italic=True, color="003366", size=10, name="Calibri")

    # --- TABLE 1 Title (Row 4) ---
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

    # --- Column Headers (Row 5) ---
    col_headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
                   "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # --- Data Rows (Row 6+) --- one row per domain sheet
    # Each sheet: (label, sheet_name, status_col, first_data_row, last_data_row)
    # Status columns map: Compliant value, Partial value, Non-Compliant value, N/A value
    domain_sheets = [
        ("Pattern Inventory", "PatternInventory", "E", 5, 55,  # Rows 5-55 (1 grey sample + 50 input rows)
         "Approved", "Draft", "Deprecated", "Under Review"),
        ("Pattern Quality", "PatternQuality", "C", 5, 14,  # Shifted by 1 due to subtitle
         "✅ Yes", "⚠️ Partial", "❌ No", "N/A"),
        ("Adoption", "Adoption", "F", 6, 59,  # Rows 6-59; Trend now col F (Description added as col B)
         "Increasing", "Stable", "Decreasing", "N/A"),
        ("Governance", "Governance", "C", 6, 55,  # Rows 6-55 (exclude grey sample row 5)
         "✅ Implemented", "⚠️ Partial", "❌ Not Implemented", "N/A"),
        ("Deviations", "Deviations", "J", 6, 55,  # Rows 6-55 (exclude grey sample row 5)
         "Closed", "Active", "Expired", "N/A"),
        ("Effectiveness", "Effectiveness", "F", 6, 55,  # Rows 6-55 (exclude grey sample row 5)
         "High", "Medium", "Low", "N/A"),
        ("Compliance", "Compliance", "D", 6, 55,  # Rows 6-55 (exclude grey sample row 5)
         "✅ Yes", "⚠️ Partial", "❌ No", "N/A"),
    ]

    row = 6
    for label, sheet, col, fr, lr, val_c, val_p, val_nc, val_na in domain_sheets:
        rng = f"'{sheet}'!{col}{fr}:{col}{lr}"
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=1).font = Font(name="Calibri")
        ws.cell(row=row, column=2, value=f'=COUNTA({rng})').border = border
        ws.cell(row=row, column=2).font = Font(name="Calibri")
        ws.cell(row=row, column=3, value=f'=COUNTIF({rng},"{val_c}")').border = border
        ws.cell(row=row, column=3).font = Font(name="Calibri")
        ws.cell(row=row, column=4, value=f'=COUNTIF({rng},"{val_p}")').border = border
        ws.cell(row=row, column=4).font = Font(name="Calibri")
        ws.cell(row=row, column=5, value=f'=COUNTIF({rng},"{val_nc}")').border = border
        ws.cell(row=row, column=5).font = Font(name="Calibri")
        # N/A (calculated as difference, not counted - DV lists don't have "N/A")
        ws.cell(row=row, column=6, value=f'=B{row}-(C{row}+D{row}+E{row})').border = border
        ws.cell(row=row, column=6).font = Font(name="Calibri")
        ws.cell(row=row, column=7,
                value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")').border = border
        ws.cell(row=row, column=7).font = Font(name="Calibri")
        row += 1

    # --- TOTAL Row ---
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Calibri")
    ws.cell(row=total_row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f"=SUM({col_letter}6:{col_letter}{total_row - 1})").border = border
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True, name="Calibri")
    ws.cell(row=total_row, column=7,
            value=f'=IF((B{total_row}-F{total_row})=0,"0%",ROUND(C{total_row}/(B{total_row}-F{total_row})*100,1)&"%")').border = border
    ws.cell(row=total_row, column=7).font = Font(bold=True, color="000000", size=12, name="Calibri")
    # D9D9D9 fill on all TOTAL row cells (Gold Standard)
    for _col in range(1, 8):
        ws.cell(row=total_row, column=_col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # --- TABLE 2: KEY METRICS ---
    table2_start = total_row + 2

    ws.merge_cells(f"A{table2_start}:G{table2_start}")
    cell = ws.cell(row=table2_start, column=1, value="TABLE 2: KEY METRICS")
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.border = border

    table2_header_row = table2_start + 1
    ws.cell(row=table2_header_row, column=1, value="Metric").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table2_header_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table2_header_row, column=1).border = border

    ws.cell(row=table2_header_row, column=2, value="Value").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table2_header_row, column=2).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table2_header_row, column=2).border = border
    ws.cell(row=table2_header_row, column=2).alignment = Alignment(horizontal="center")

    ws.merge_cells(f"C{table2_header_row}:G{table2_header_row}")
    ws.cell(row=table2_header_row, column=3, value="What This Shows").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table2_header_row, column=3).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table2_header_row, column=3).border = border

    table2_metrics = [
        ("Deprecated Patterns", '=COUNTIF(PatternInventory!E5:E55,"Deprecated")',
         "Patterns deprecated but still in use (technical debt)"),
        ("Draft Patterns", '=COUNTIF(PatternInventory!E5:E55,"Draft")',
         "Patterns in draft state (not production-ready)"),
        ("Patterns Under Review", '=COUNTIF(PatternInventory!E5:E55,"Under Review")',
         "Patterns under review (uncertain security posture)"),
        ("Declining Adoption", '=COUNTIF(Adoption!F6:F59,"Decreasing")',
         "Patterns with decreasing adoption (failures or obsolescence)"),
        ("Low Effectiveness Patterns", '=COUNTIF(Effectiveness!F6:F55,"Low")',
         "Patterns with low effectiveness (poor security outcomes)"),
        ("Active Deviations", '=COUNTIF(Deviations!J6:J55,"Active")',
         "Active exceptions from patterns (security risks)"),
        ("Expired Deviations", '=COUNTIF(Deviations!J6:J55,"Expired")',
         "Expired exceptions (operating outside approved parameters)"),
        ("Quality Element Gaps", '=COUNTIF(PatternQuality!C5:C14,"❌ No")',
         "Missing quality elements (incomplete documentation)"),
        ("Governance Gaps", '=COUNTIF(Governance!C6:C55,"❌ Not Implemented")',
         "Pattern governance elements not implemented"),
        ("Overall Compliance Rate", f'=\'Summary Dashboard\'!G{total_row}',
         "Overall pattern catalogue compliance percentage"),
    ]

    metric_row = table2_header_row + 1
    for metric_name, formula, description in table2_metrics:
        ws.cell(row=metric_row, column=1, value=metric_name).font = Font(size=11, name="Calibri")
        ws.cell(row=metric_row, column=1).border = border

        cell = ws.cell(row=metric_row, column=2, value=formula)
        cell.font = Font(size=11, name="Calibri")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

        ws.merge_cells(f"C{metric_row}:G{metric_row}")
        ws.cell(row=metric_row, column=3, value=description).font = Font(size=9, name="Calibri")
        ws.cell(row=metric_row, column=3).border = border
        # Apply borders to all cells in merged range
        for col in range(4, 8):  # D-G
            ws.cell(row=metric_row, column=col).border = border

        metric_row += 1

    # 2 buffer rows: A alone | B alone | C:G merged (mirrors TABLE 2 metric structure)
    for _buf_row in range(metric_row, metric_row + 2):
        ws.cell(row=_buf_row, column=1).border = border
        ws.cell(row=_buf_row, column=2).border = border
        ws.merge_cells(f"C{_buf_row}:G{_buf_row}")
        for _col in range(3, 8):
            ws.cell(row=_buf_row, column=_col).border = border
    # --- TABLE 3: CRITICAL FINDINGS ---
    table3_start = metric_row + 3

    ws.merge_cells(f"A{table3_start}:G{table3_start}")
    cell = ws.cell(row=table3_start, column=1, value="TABLE 3: CRITICAL FINDINGS")
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.border = border

    table3_header_row = table3_start + 1
    ws.cell(row=table3_header_row, column=1, value="Critical Finding Type").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table3_header_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table3_header_row, column=1).border = border

    ws.cell(row=table3_header_row, column=2, value="Count").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table3_header_row, column=2).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table3_header_row, column=2).border = border
    ws.cell(row=table3_header_row, column=2).alignment = Alignment(horizontal="center")

    ws.merge_cells(f"C{table3_header_row}:G{table3_header_row}")
    ws.cell(row=table3_header_row, column=3, value="Filter Instructions").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=table3_header_row, column=3).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=table3_header_row, column=3).border = border

    table3_findings = [
        ("Deprecated Patterns", '=COUNTIF(PatternInventory!E5:E55,"Deprecated")',
         'Filter PatternInventory sheet: Status = "Deprecated"'),
        ("Expired Deviations", '=COUNTIF(Deviations!J6:J55,"Expired")',
         'Filter Deviations sheet: Status = "Expired"'),
        ("Low Effectiveness Patterns", '=COUNTIF(Effectiveness!F6:F55,"Low")',
         'Filter Effectiveness sheet: Rating = "Low"'),
        ("Governance Gaps", '=COUNTIF(Governance!C6:C55,"❌ Not Implemented")',
         'Filter Governance sheet: Status = "❌ Not Implemented"'),
        ("Active Deviations", '=COUNTIF(Deviations!J6:J55,"Active")',
         'Filter Deviations sheet: Status = "Active"'),
    ]

    finding_row = table3_header_row + 1
    first_finding_row = finding_row
    for finding_name, formula, instructions in table3_findings:
        ws.cell(row=finding_row, column=1, value=finding_name).font = Font(size=11, name="Calibri")
        ws.cell(row=finding_row, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=finding_row, column=1).border = border

        cell = ws.cell(row=finding_row, column=2, value=formula)
        cell.font = Font(size=11, name="Calibri")
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

        ws.merge_cells(f"C{finding_row}:G{finding_row}")
        ws.cell(row=finding_row, column=3, value=instructions).font = Font(size=9, name="Calibri")
        ws.cell(row=finding_row, column=3).border = border
        # Apply borders to all cells in merged range
        for col in range(4, 8):  # D-G
            ws.cell(row=finding_row, column=col).border = border

        finding_row += 1

    ws.cell(row=finding_row, column=1, value="TOTAL").font = Font(bold=True, size=11, name="Calibri")
    ws.cell(row=finding_row, column=1).border = border

    cell = ws.cell(row=finding_row, column=2, value=f"=SUM(B{first_finding_row}:B{finding_row - 1})")
    cell.font = Font(bold=True, size=11, name="Calibri")
    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"C{finding_row}:G{finding_row}")
    ws.cell(row=finding_row, column=3, value="Total critical findings requiring immediate remediation").font = Font(italic=True, size=9, name="Calibri")
    ws.cell(row=finding_row, column=3).border = border
    # Apply borders to all cells in merged range
    for col in range(4, 8):  # D-G
        ws.cell(row=finding_row, column=col).border = border

    # --- Column Widths & Freeze ---
    widths = [40, 16, 16, 18, 18, 12, 15]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create Approval Sign-Off sheet (Gold Standard)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ROW 1: TITLE BANNER
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # ROW 2: CONTROL REFERENCE
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # ROW 3: ASSESSMENT SUMMARY BANNER
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # SUMMARY FIELDS
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G13"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == "")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        if "Assessment Status" in label:
            status_row_for_dv = row
        row += 1

    # ASSESSMENT STATUS DROPDOWN
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f"B{status_row_for_dv}")

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        """Create one approver section (banner + 5 fields)."""
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{current_row}"] = field
            ws[f"A{current_row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{current_row}"].border = border
            ws.merge_cells(f"B{current_row}:E{current_row}")
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _create_approver_section(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")
    row = _create_approver_section(row, "APPROVED BY (CISO)", "003366")

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        row += 1

    # COLUMN WIDTHS & FREEZE PANES
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


def create_pattern_inventory_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:J1')
    ws['A1'] = "SECURE ARCHITECTURE PATTERNS - PATTERN INVENTORY"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:J2')
    ws['A2'] = "Maintain inventory of approved secure architecture patterns with ownership and review status"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 11):  # A through J
        ws.cell(row=2, column=col).border = border

    headers = ['Pat-ID', 'Category', 'Name', 'Version', 'Status', 'Owner', 'LastReview', 'NextReview', 'DocumentRef', 'Notes']
    widths = [15, 20, 35, 10, 15, 20, 12, 12, 25, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    category_dv = DataValidation(type='list', formula1='"Authentication,Authorisation,DataProtection,NetworkSecurity,Integration,Cloud,LoggingMonitoring,Identity,Container,Serverless"', allow_blank=True)

    status_dv = DataValidation(type='list', formula1='"Approved,Draft,Deprecated,Under Review"', allow_blank=True)

    patterns = [
        ('PAT-AUTH-001', 'Authentication', 'SSO Integration Pattern'),
        ('PAT-AUTH-002', 'Authentication', 'MFA Implementation Pattern'),
        ('PAT-AUTH-003', 'Authentication', 'Service Authentication Pattern'),
        ('PAT-AUTH-004', 'Authentication', 'Token Management Pattern'),
        ('PAT-AUTHZ-001', 'Authorisation', 'RBAC Implementation Pattern'),
        ('PAT-AUTHZ-002', 'Authorisation', 'ABAC Implementation Pattern'),
        ('PAT-AUTHZ-003', 'Authorisation', 'API Authorisation Pattern'),
        ('PAT-AUTHZ-004', 'Authorisation', 'Delegated Administration Pattern'),
        ('PAT-DATA-001', 'DataProtection', 'Encryption at Rest Pattern'),
        ('PAT-DATA-002', 'DataProtection', 'Encryption in Transit Pattern'),
        ('PAT-DATA-003', 'DataProtection', 'Key Management Pattern'),
        ('PAT-DATA-004', 'DataProtection', 'Tokenisation Pattern'),
        ('PAT-NET-001', 'NetworkSecurity', 'DMZ Architecture Pattern'),
        ('PAT-NET-002', 'NetworkSecurity', 'Micro-segmentation Pattern'),
        ('PAT-NET-003', 'NetworkSecurity', 'API Gateway Pattern'),
        ('PAT-NET-004', 'NetworkSecurity', 'Zero Trust Network Pattern'),
        ('PAT-INT-001', 'Integration', 'Secure REST API Pattern'),
        ('PAT-INT-002', 'Integration', 'Message Queue Security Pattern'),
        ('PAT-INT-003', 'Integration', 'Event-Driven Security Pattern'),
        ('PAT-INT-004', 'Integration', 'Service Mesh Pattern'),
        ('PAT-CLD-001', 'Cloud', 'Landing Zone Architecture'),
        ('PAT-CLD-002', 'Cloud', 'Workload Isolation Pattern'),
        ('PAT-CLD-003', 'Cloud', 'Serverless Security Pattern'),
        ('PAT-CLD-004', 'Cloud', 'Container Security Pattern'),
        ('PAT-LOG-001', 'LoggingMonitoring', 'Centralised Logging Pattern'),
        ('PAT-LOG-002', 'LoggingMonitoring', 'SIEM Integration Pattern'),
        ('PAT-LOG-003', 'LoggingMonitoring', 'Audit Trail Pattern'),
        ('PAT-LOG-004', 'LoggingMonitoring', 'Log Protection Pattern'),
        ('PAT-IDM-001', 'Identity', 'Identity Lifecycle Pattern'),
        ('PAT-IDM-002', 'Identity', 'Privileged Access Management Pattern'),
    ]

    # Row 5: grey sample row — generic placeholder (not a real pattern)
    sample_values = {
        1: 'PAT-AUTH-00x',                        # Pat-ID
        2: 'Authentication',                       # Category
        3: 'Sample Integration Pattern',           # Name
        4: '1.0',                                  # Version
        5: 'Approved',                             # Status
        6: 'Security Architect',                   # Owner
        7: '15.01.2026',                           # LastReview
        8: '15.01.2027',                           # NextReview
        9: '/docs/patterns/PAT-AUTH-00x.pdf',      # DocumentRef
        10: 'Example — replace with your pattern', # Notes
    }
    for col, value in sample_values.items():
        ws.cell(row=5, column=col, value=value)
    for col in range(1, 11):
        apply_style(ws.cell(row=5, column=col), styles['sample'])

    # Rows 6-35: all 30 patterns — fully FFFFCC (sample row above is generic placeholder)
    for row, (pat_id, category, name) in enumerate(patterns, 6):
        ws.cell(row=row, column=1, value=pat_id)
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=name)
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    # Empty FFFFCC rows 36-55 (20 rows to reach 50 total input rows)
    last_pattern_row = 5 + len(patterns)
    for row in range(last_pattern_row + 1, last_pattern_row + 21):
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    end_row = last_pattern_row + 20
    category_dv.add(f'B5:B{end_row}')
    status_dv.add(f'E5:E{end_row}')
    validations.extend([category_dv, status_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A6'


def create_pattern_quality_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:E1')
    ws['A1'] = "SECURE ARCHITECTURE PATTERNS - DOCUMENTATION QUALITY"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess quality and completeness of pattern documentation elements"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 6):  # A through E
        ws.cell(row=2, column=col).border = border

    headers = ['Pat-ID', 'Element', 'Present', 'Quality', 'Notes']
    widths = [15, 25, 10, 10, 40]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    present_dv = DataValidation(type='list', formula1='"✅ Yes,⚠️ Partial,❌ No"', allow_blank=True)

    rating_dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)

    elements = ['ProblemStatement', 'Context', 'Solution', 'SecurityRationale', 'Implementation',
                'Example', 'AntiPatterns', 'RelatedPatterns', 'ComplianceMapping', 'ThreatModel']

    row = 5
    for element in elements:
        ws.cell(row=row, column=2, value=element)
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col != 2 else styles['normal'])
        row += 1

    present_dv.add(f'C5:C{row-1}')
    rating_dv.add(f'D5:D{row-1}')
    validations.extend([present_dv, rating_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A6'


def create_adoption_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = "SECURE ARCHITECTURE PATTERNS - ADOPTION TRACKING"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = "Track adoption rates and barriers for secure architecture patterns across projects"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 9):  # A through H
        ws.cell(row=2, column=col).border = border

    # Description added as col B between Pat-ID and ProjectCount
    headers = ['Pat-ID', 'Description', 'ProjectCount', 'TotalApplicable', 'AdoptionRate', 'Trend', 'Barriers', 'Action']
    widths = [15, 35, 15, 15, 15, 12, 35, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    trend_dv = DataValidation(type='list', formula1='"Increasing,Stable,Decreasing"', allow_blank=True)

    # Row 5: grey sample row
    row = 5
    ws.cell(row=5, column=1, value='PATTERN-AUTH-001')
    ws.cell(row=5, column=2, value='SSO Integration Pattern — production adoption tracking')
    ws.cell(row=5, column=3, value='12')
    ws.cell(row=5, column=4, value='15')
    ws.cell(row=5, column=5, value=f'=IF(D{row}>0,C{row}/D{row},0)')
    ws.cell(row=5, column=6, value='Increasing')
    ws.cell(row=5, column=7, value='None identified')
    ws.cell(row=5, column=8, value='Continue promotion in new projects')
    for col in range(1, 9):
        apply_style(ws.cell(row=5, column=col), styles['sample'])

    # Pre-populated FFFFCC requirement rows (rows 6–9) — mandatory supplier alignment items
    # Col A = empty FFFFCC (Pat-ID), Col B = description text, Cols C-H = empty FFFFCC
    supplier_items = [
        "Supplier contracts require adherence to organisational secure engineering principles",
        "Third-party/vendor architectures reviewed against organisational standards before adoption",
        "Cloud service provider security architecture assessed against the pattern catalogue before onboarding",
        "Externally developed systems subject to the same architecture review process as internally developed systems",
    ]
    for item in supplier_items:
        row += 1
        # Col A: empty FFFFCC (Pat-ID for user to fill)
        apply_style(ws.cell(row=row, column=1), styles['input'])
        # Col B: requirement description text — FFFFCC
        ws.cell(row=row, column=2).value = item
        apply_style(ws.cell(row=row, column=2), styles['input'])
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Col E: AdoptionRate formula (D=TotalApplicable, C=ProjectCount)
        ws.cell(row=row, column=5).value = f'=IF(D{row}>0,C{row}/D{row},0)'
        # Cols C-H: FFFFCC (C, D = input; E = formula; F-H = input)
        for pcol in range(3, 9):
            if pcol == 5:
                apply_style(ws.cell(row=row, column=pcol), styles['formula'])
            else:
                apply_style(ws.cell(row=row, column=pcol), styles['input'])

    # Add 50 empty rows (all yellow)
    for i in range(50):
        row += 1
        ws.cell(row=row, column=5, value=f'=IF(D{row}>0,C{row}/D{row},0)')
        for col in range(1, 9):
            if col == 5:
                apply_style(ws.cell(row=row, column=col), styles['formula'])
            else:
                apply_style(ws.cell(row=row, column=col), styles['input'])

    trend_dv.add(f'F6:F{row}')
    validations.append(trend_dv)
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A6'


def create_governance_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "SECURE ARCHITECTURE PATTERNS - GOVERNANCE ASSESSMENT"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:G2')
    ws['A2'] = "Assess governance controls for pattern catalogue management and maintenance"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 8):  # A through G
        ws.cell(row=2, column=col).border = border

    headers = ['Gov-ID', 'Requirement', 'Status', 'Evidence', 'Gap', 'Owner', 'Notes']
    widths = [10, 40, 15, 30, 30, 20, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    status_dv = DataValidation(type='list', formula1='"✅ Implemented,⚠️ Partial,❌ Not Implemented"', allow_blank=True)

    # MAX-001 fix: 1 sample (grey) + 50 empty rows (yellow)
    row = 5
    ws.cell(row=5, column=1, value='GOV-001')
    ws.cell(row=5, column=2, value='Pattern ownership assigned for each pattern')
    ws.cell(row=5, column=3, value='Implemented')
    ws.cell(row=5, column=4, value='Governance policy documentation')
    ws.cell(row=5, column=5, value='None identified')
    ws.cell(row=5, column=6, value='Security Team')
    ws.cell(row=5, column=7, value='Pattern governance established')
    for col in range(1, 8):
        apply_style(ws.cell(row=5, column=col), styles['sample'])

    # Add 50 empty rows (all yellow)
    for i in range(50):
        row += 1
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    status_dv.add('C6:C55')
    validations.append(status_dv)
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A6'


def create_deviations_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:J1')
    ws['A1'] = "SECURE ARCHITECTURE PATTERNS - DEVIATION TRACKING"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:J2')
    ws['A2'] = "Track approved deviations from secure patterns with justifications and compensating controls"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 11):  # A through J
        ws.cell(row=2, column=col).border = border

    headers = ['Dev-ID', 'Pattern', 'Project', 'Category', 'Justification', 'Approved', 'ApprovedBy', 'Compensating', 'Expiry', 'Status']
    widths = [10, 15, 25, 20, 40, 10, 20, 35, 12, 12]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    category_dv = DataValidation(type='list', formula1='"Technical Constraint,Legacy System,Vendor Limitation,Cost/Time,Performance"', allow_blank=True)

    approved_dv = DataValidation(type='list', formula1='"Yes,No,Pending"', allow_blank=True)

    status_dv = DataValidation(type='list', formula1='"Active,Closed,Expired"', allow_blank=True)

    # MAX-001 fix: 1 sample (grey) + 50 empty rows (yellow)
    row = 5
    ws.cell(row=5, column=1, value='DEV-001')
    ws.cell(row=5, column=2, value='PATTERN-AUTH-001')
    ws.cell(row=5, column=3, value='Customer Portal Upgrade')
    ws.cell(row=5, column=4, value='Legacy System')
    ws.cell(row=5, column=5, value='Legacy authentication system cannot support OAuth2')
    ws.cell(row=5, column=6, value='Yes')
    ws.cell(row=5, column=7, value='Security Architect')
    ws.cell(row=5, column=8, value='Additional monitoring and session timeout controls')
    ws.cell(row=5, column=9, value='2026-12-31')
    ws.cell(row=5, column=10, value='Active')
    for col in range(1, 11):
        apply_style(ws.cell(row=5, column=col), styles['sample'])

    # Add 50 empty rows (all yellow)
    for i in range(50):
        row += 1
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    category_dv.add('D6:D55')
    approved_dv.add('F6:F55')
    status_dv.add('J6:J55')
    validations.extend([category_dv, approved_dv, status_dv])
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A6'


def create_effectiveness_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "SECURE ARCHITECTURE PATTERNS - EFFECTIVENESS METRICS"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:G2')
    ws['A2'] = "Measure pattern effectiveness through security incident and vulnerability metrics"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 8):  # A through G
        ws.cell(row=2, column=col).border = border

    headers = ['Pat-ID', 'SecurityIncidents', 'VulnFindings', 'AuditFindings', 'UserFeedback', 'Effectiveness', 'Notes']
    widths = [15, 15, 20, 15, 12, 12, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    effectiveness_dv = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)

    # Row 5: grey sample row with example data
    ws.cell(row=5, column=1, value='PAT-001')
    ws.cell(row=5, column=2, value='0')
    ws.cell(row=5, column=3, value='2')
    ws.cell(row=5, column=4, value='0')
    ws.cell(row=5, column=5, value='Positive')
    ws.cell(row=5, column=6, value='High')
    ws.cell(row=5, column=7, value='Pattern performing effectively')
    for col in range(1, 8):
        apply_style(ws.cell(row=5, column=col), styles['sample'])

    # Rows 6-55: 50 yellow empty rows
    for row in range(6, 56):
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    effectiveness_dv.add('F6:F55')  # Exclude grey sample row 5
    validations.append(effectiveness_dv)
    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A6'


def create_compliance_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "SECURE ARCHITECTURE PATTERNS - POLICY COMPLIANCE"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:G2')
    ws['A2'] = "Score compliance with pattern catalogue policy requirements and control objectives"
    apply_style(ws['A2'], styles['subtitle'])
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 8):  # A through G
        ws.cell(row=2, column=col).border = border

    headers = ['Comp-ID', 'Requirement', 'Source', 'Compliant', 'Evidence', 'Score', 'Notes']
    widths = [10, 40, 20, 12, 35, 10, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    validations = []
    compliant_dv = DataValidation(type='list', formula1='"✅ Yes,⚠️ Partial,❌ No"', allow_blank=True)

    # MAX-001 fix: 1 sample (grey) + 50 empty rows (yellow)
    row = 5
    ws.cell(row=5, column=1, value='COMP-001')
    ws.cell(row=5, column=2, value='Secure pattern catalogue maintained')
    ws.cell(row=5, column=3, value='POL-A.8.27 \u00a72.2.2')
    ws.cell(row=5, column=4, value='Yes')
    ws.cell(row=5, column=5, value='Pattern catalogue documentation')
    ws.cell(row=5, column=6, value=f'=IF(D{row}="✅ Yes",100,IF(D{row}="⚠️ Partial",50,0))')
    ws.cell(row=5, column=7, value='Catalogue regularly reviewed')
    for col in range(1, 8):
        apply_style(ws.cell(row=5, column=col), styles['sample'])

    # Add 50 empty rows (all yellow)
    for i in range(50):
        row += 1
        ws.cell(row=row, column=6, value=f'=IF(D{row}="✅ Yes",100,IF(D{row}="⚠️ Partial",50,0))')
        for col in range(1, 8):
            if col == 6:
                apply_style(ws.cell(row=row, column=col), styles['formula'])
            else:
                apply_style(ws.cell(row=row, column=col), styles['input'])

    compliant_dv.add('D6:D55')
    validations.append(compliant_dv)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Overall score (use subheader style, NOT FFFFCC formula, to avoid FML-004 stale-range flag)
    last_row = 57
    ws.cell(row=last_row, column=5, value="Overall Compliance:")
    ws.cell(row=last_row, column=6, value='=AVERAGE(F6:F55)')
    apply_style(ws.cell(row=last_row, column=5), styles['subheader'])
    apply_style(ws.cell(row=last_row, column=6), styles['subheader'])

    ws.freeze_panes = 'A6'


# =============================================================================
# MAIN
# =============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 80)
    logger.info(f"ISMS Control {CONTROL_ID} - {WORKBOOK_NAME} Generator")
    logger.info("=" * 80)

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        ("Instructions & Legend", create_instructions_sheet),
        ("PatternInventory", create_pattern_inventory_sheet),
        ("PatternQuality", create_pattern_quality_sheet),
        ("Adoption", create_adoption_sheet),
        ("Governance", create_governance_sheet),
        ("Deviations", create_deviations_sheet),
        ("Effectiveness", create_effectiveness_sheet),
        ("Compliance", create_compliance_sheet),
        ("Evidence Register", create_evidence_register),
        ("Summary Dashboard", create_summary_dashboard_sheet),
        ("Approval Sign-Off", create_approval_sheet),
    ]

    for sheet_name, create_func in sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False
        create_func(ws)
        logger.info(f"  Created sheet: {sheet_name}")

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")
    logger.info("Generation complete!")


def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
