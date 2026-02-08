#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.28.3 - Code Review & Testing Assessment
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Assessment Domain 3 of 4: Code Review and Security Testing Processes

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific code review practices, testing methodologies,
and quality assurance processes.

Key customization areas:
1. Code review workflows and tools (match your review process and platforms)
2. Security testing types and coverage (adapt to your testing strategy)
3. Testing tools and frameworks (specific to your technology stack)
4. Review criteria and quality gates (based on your quality standards)
5. Compliance thresholds and scoring (aligned with your risk tolerance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
code review effectiveness and security testing coverage across all development
projects and repositories.

**Purpose:**
Enables systematic assessment of code review and security testing processes
against ISO 27001:2022 Control A.8.28 requirements, supporting evidence-based
validation of quality assurance and vulnerability detection practices.

**Assessment Scope:**
- Peer code review processes and coverage metrics
- Security-focused code review checklists and criteria
- Static Application Security Testing (SAST) integration and coverage
- Dynamic Application Security Testing (DAST) implementation
- Software Composition Analysis (SCA) for dependencies
- Interactive Application Security Testing (IAST) where applicable
- Penetration testing frequency and scope
- Security regression testing and test case management
- Code coverage metrics and security test coverage
- Vulnerability management and remediation tracking
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and testing standard references
2. Peer Review - Code review process effectiveness and coverage assessment
3. Security Review - Security-focused review criteria and execution
4. SAST Coverage - Static analysis implementation and finding resolution
5. DAST Testing - Dynamic testing coverage and vulnerability detection
6. SCA Analysis - Dependency scanning and vulnerable component management
7. Penetration Testing - Pen test frequency, scope, and finding remediation
8. Test Coverage - Code coverage and security test case adequacy
9. Vulnerability Mgmt - Defect tracking, prioritization, and remediation SLAs
10. Gap Analysis - Inadequate review/testing coverage and remediation
11. Evidence Register - Audit evidence tracking and documentation
12. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with testing methodology dropdown lists
- Conditional formatting for coverage thresholds and compliance status
- Automated gap identification for untested code or missing test types
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with test management and defect tracking systems

**Integration:**
This assessment feeds into the A.8.28.5 Compliance Dashboard, which
consolidates data from all four assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a828_3_code_review_testing.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a828_3_code_review_testing.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a828_3_code_review_testing.py --date 20250124

Output:
    File: ISMS_A_8_28_3_Code_Review_Testing_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize testing standards to match your QA methodology
    2. Inventory all code repositories, projects, and applications
    3. Complete assessments for each project or development team
    4. Validate SAST/DAST/SCA tool coverage and scan frequency
    5. Review penetration testing scope and remediation timelines
    6. Conduct gap analysis for insufficient testing or review coverage
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (scan reports, review records, pen test results)
    9. Obtain stakeholder approvals (QA Lead, AppSec, Dev Manager)
    10. Feed results into A.8.28.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    3 of 4 (Code Review & Testing)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Implementation Guide
    - ISMS-IMP-A.8.28.1: SDLC Integration Assessment (Domain 1)
    - ISMS-IMP-A.8.28.2: Standards & Tools Assessment (Domain 2)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Assessment (Domain 4)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a828_1_sdlc_assessment.py
    - generate_a828_2_standards_tools.py
    - generate_a828_4_third_party_oss.py
    - generate_a828_5_compliance_dashboard.py
    - normalize_assessment_files_a828.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.28.3 specification
    - Supports comprehensive code review and security testing evaluation
    - Integrated with A.8.28.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Testing Strategy Balance:**
Effective application security requires layered testing approaches. Don't
rely solely on automated tools (SAST/DAST/SCA) - they miss context-specific
vulnerabilities. Don't rely solely on manual testing - it doesn't scale.
Balance is key.

**Code Review Effectiveness:**
Code review participation rates don't equal effectiveness. A rushed review
that approves everything is worse than no review. Assess quality of reviews:
- Are security issues actually caught in review?
- How long do reviews take on average?
- What percentage result in meaningful feedback vs. rubber-stamping?

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of testing coverage, review records, and
vulnerability remediation timelines.

**Data Protection:**
Assessment workbooks contain sensitive security information including:
- Vulnerability details and security defect patterns
- Penetration test findings and attack scenarios
- Code coverage gaps and untested functionality
- Security testing tool configurations and results

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Review testing coverage and vulnerability remediation rates
- Semi-annually: Update testing methodologies for new threat patterns
- Annually: Complete reassessment of all projects and applications
- Ad-hoc: When new testing tools adopted or testing requirements change

**Quality Assurance:**
Have application security SMEs, QA leads, and security testing specialists
validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
Ensure testing standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 testing requirements (Requirement 6.3, 11.3)
- Healthcare: HIPAA security testing and vulnerability management
- Finance: Regional banking penetration testing requirements
- Government: Jurisdiction-specific security testing mandates

Customize assessment criteria to include regulatory-specific requirements.

**Integration with Testing Tools:**
Where possible, integrate assessment data with:
- Code review platforms (GitHub, GitLab, Bitbucket, Gerrit, etc.)
- Test management systems (Jira, Azure DevOps, TestRail, etc.)
- SAST/DAST/SCA tool outputs (detailed scan results and trends)
- Defect tracking systems (vulnerability lifecycle and remediation metrics)
- CI/CD pipelines (automated test execution and coverage reports)

Automated tool integration reduces manual effort and improves accuracy.

**Don't Fool Yourself - Feynman Principle:**
High test coverage percentages don't automatically mean good security testing.
100% code coverage with zero security test cases is worthless. Similarly,
running SAST scans doesn't help if findings aren't reviewed and addressed.

Assess actual security defect detection and remediation:
- How many security vulnerabilities were found in the last 6 months?
- What percentage were found by automated vs. manual testing?
- How quickly are high/critical findings remediated?
- Can you show examples of security issues caught before production?

If you can't demonstrate defects found and fixed, your testing isn't working.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.28.3"
WORKBOOK_NAME = "Code Review and Security Testing Processes"
CONTROL_ID = "A.8.28"
CONTROL_NAME = "Secure Coding"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

import sys


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.28.3 specification
    sheets = [
        "Instructions",
        "Code_Review_Process",
        "Security_Champion_Review",
        "Unit_Integration_Testing",
        "API_Application_Testing",
        "External_Testing_Validation",
        "Summary_Dashboard",
        "Evidence_Register",
        "Gap_Analysis",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Code review and testing are the last lines of defense before production.
    This workbook helps assess whether those defenses are actually working.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_implemented": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notimplemented": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        },
        "priority_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "priority_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "priority_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "priority_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell with fresh objects."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    
    Process without effectiveness is theater. These validations help
    assess whether reviews and tests actually catch vulnerabilities.
    """
    validations = {
        # Implementation status
        'implementation_status': DataValidation(
            type="list",
            formula1='"Implemented,Partially Implemented,Not Implemented,N/A"',
            allow_blank=False
        ),
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Unknown"',
            allow_blank=False
        ),
        
        # Review and testing effectiveness
        'effectiveness_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Fair,Poor,Not Assessed"',
            allow_blank=False
        ),
        'coverage_level': DataValidation(
            type="list",
            formula1='"Complete (>95%),High (80-95%),Medium (60-79%),Low (<60%),Unknown"',
            allow_blank=False
        ),
        'frequency': DataValidation(
            type="list",
            formula1='"Every Change,Daily,Weekly,Monthly,Quarterly,On-Demand,Never"',
            allow_blank=False
        ),
        
        # Priority and risk
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Negligible"',
            allow_blank=False
        ),
        
        # Gap status
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Closed,Deferred"',
            allow_blank=False
        ),
        
        # Evidence types
        'evidence_type': DataValidation(
            type="list",
            formula1='"Review Log,Test Result,Pentest Report,Bug Bounty Submission,Training Record,Policy Document,Tool Output,Screenshot,Other"',
            allow_blank=False
        ),
        
        # Review types
        'review_type': DataValidation(
            type="list",
            formula1='"Peer Review,Security Champion Review,Architecture Review,Design Review,Post-Incident Review"',
            allow_blank=False
        ),
        
        # Test types
        'test_type': DataValidation(
            type="list",
            formula1='"Unit Test,Integration Test,API Test,Security Test,Penetration Test,Regression Test,Smoke Test"',
            allow_blank=False
        ),
        
        # Test status
        'test_status': DataValidation(
            type="list",
            formula1='"Passing,Failing,Flaky,Disabled,Not Run"',
            allow_blank=False
        ),
        
        # Automation level
        'automation_level': DataValidation(
            type="list",
            formula1='"Fully Automated,Partially Automated,Manual,Not Applicable"',
            allow_blank=False
        ),
        
        # Finding severity
        'severity': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Informational"',
            allow_blank=False
        ),
        
        # Remediation status
        'remediation_status': DataValidation(
            type="list",
            formula1='"Fixed,In Progress,Accepted Risk,False Positive,Deferred"',
            allow_blank=False
        ),
        
        # Compliance level
        'compliance_level': DataValidation(
            type="list",
            formula1='"Fully Compliant,Mostly Compliant,Partially Compliant,Non-Compliant,N/A"',
            allow_blank=False
        ),
        
        # Training status
        'training_status': DataValidation(
            type="list",
            formula1='"Completed,In Progress,Not Started,N/A"',
            allow_blank=False
        ),
        
        # Approval decision
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


# ============================================================================
# SECTION 3: ASSESSMENT DOMAIN DATA DEFINITIONS
# ============================================================================

def get_domain1_requirements():
    """
    Domain 1: Code Review Process & Governance
    
    Code review is only effective if it's mandatory, done by qualified people,
    and actually catches security issues. "LGTM" is not a security review.
    """
    return [
        {
            "id": "1.1",
            "requirement": "Peer review is mandatory for all production code changes",
            "evidence": "Repository settings showing branch protection, review approval requirements",
            "category": "Review Policy"
        },
        {
            "id": "1.2",
            "requirement": "Code review policy is documented and accessible to all developers",
            "evidence": "Code review policy document, developer handbook, internal wiki",
            "category": "Review Policy"
        },
        {
            "id": "1.3",
            "requirement": "Minimum reviewer qualifications are defined and enforced",
            "evidence": "Reviewer qualification criteria, approved reviewer list",
            "category": "Review Policy"
        },
        {
            "id": "1.4",
            "requirement": "All reviewers are trained on secure code review practices",
            "evidence": "Training completion records, secure code review training materials",
            "category": "Reviewer Training"
        },
        {
            "id": "1.5",
            "requirement": "Security-focused checklist is used in all code reviews",
            "evidence": "Code review checklist template, checklist usage in PR comments",
            "category": "Review Process"
        },
        {
            "id": "1.6",
            "requirement": "Code review tools are integrated into development workflow",
            "evidence": "GitHub/GitLab/Bitbucket PR process, review tool configuration",
            "category": "Review Tools"
        },
        {
            "id": "1.7",
            "requirement": "Review approval is required before code can be merged to production",
            "evidence": "Branch protection rules, merge approval logs",
            "category": "Review Enforcement"
        },
        {
            "id": "1.8",
            "requirement": "Code review coverage is measured and tracked",
            "evidence": "Review coverage metrics dashboard, coverage reports",
            "category": "Review Metrics"
        },
        {
            "id": "1.9",
            "requirement": "Review comments reference secure coding standards and best practices",
            "evidence": "Sample review comments with standards references",
            "category": "Review Quality"
        },
        {
            "id": "1.10",
            "requirement": "High-risk code changes require additional security review scrutiny",
            "evidence": "Risk-based review policy, high-risk change review records",
            "category": "Review Process"
        },
        {
            "id": "1.11",
            "requirement": "Security findings from code review are tracked to resolution",
            "evidence": "Finding tracking system (JIRA, GitHub Issues), resolution logs",
            "category": "Review Follow-up"
        },
        {
            "id": "1.12",
            "requirement": "Code review metrics are reported to management quarterly",
            "evidence": "Quarterly review metrics reports, management dashboards",
            "category": "Review Metrics"
        },
        {
            "id": "1.13",
            "requirement": "Review bypass requires documented approval and justification",
            "evidence": "Exception request process, bypass approval records",
            "category": "Review Enforcement"
        },
        {
            "id": "1.14",
            "requirement": "Automated review assistance tools are deployed (PR bots, automated checks)",
            "evidence": "PR bot integration (Danger, CodeOwners), automated check configuration",
            "category": "Review Tools"
        },
        {
            "id": "1.15",
            "requirement": "Code review quality is audited periodically",
            "evidence": "Review quality audit results, sample review assessments",
            "category": "Review Quality"
        },
        {
            "id": "1.16",
            "requirement": "Self-review is not permitted for production code",
            "evidence": "Review policy prohibiting self-approval, enforcement evidence",
            "category": "Review Policy"
        },
        {
            "id": "1.17",
            "requirement": "Code review turnaround time is measured and within SLA",
            "evidence": "Review SLA metrics, turnaround time tracking",
            "category": "Review Metrics"
        },
        {
            "id": "1.18",
            "requirement": "Code review process is continuously improved based on lessons learned",
            "evidence": "Process improvement backlog, retrospective notes",
            "category": "Continuous Improvement"
        },
    ]


def get_domain2_requirements():
    """
    Domain 2: Security Champion Review & Architecture Review
    
    Security Champions are developers who become security force multipliers.
    They provide deeper security expertise where needed.
    """
    return [
        {
            "id": "2.1",
            "requirement": "Security Champion program is formally established",
            "evidence": "Program charter, Security Champion roster, program documentation",
            "category": "Champion Program"
        },
        {
            "id": "2.2",
            "requirement": "Security Champions are assigned to all development teams",
            "evidence": "Team assignments, Security Champion coverage matrix",
            "category": "Champion Program"
        },
        {
            "id": "2.3",
            "requirement": "Security Champions are trained on secure development and threat modeling",
            "evidence": "Training certifications, secure development course completion records",
            "category": "Champion Training"
        },
        {
            "id": "2.4",
            "requirement": "Criteria for Security Champion review are clearly defined",
            "evidence": "Review trigger criteria documentation (e.g., auth changes, crypto, PII)",
            "category": "Champion Process"
        },
        {
            "id": "2.5",
            "requirement": "High-risk code changes are automatically escalated to Security Champions",
            "evidence": "Escalation logs, Champion review records, automated escalation rules",
            "category": "Champion Process"
        },
        {
            "id": "2.6",
            "requirement": "Architecture review is required for all new systems and major changes",
            "evidence": "Architecture review policy, architecture review board meeting notes",
            "category": "Architecture Review"
        },
        {
            "id": "2.7",
            "requirement": "Security design review is performed before implementation begins",
            "evidence": "Design review meeting notes, security design sign-off records",
            "category": "Architecture Review"
        },
        {
            "id": "2.8",
            "requirement": "Threat modeling is performed for security-sensitive features",
            "evidence": "Threat model documents, STRIDE analysis, attack trees",
            "category": "Threat Modeling"
        },
        {
            "id": "2.9",
            "requirement": "Security Champion availability and response time are tracked and meet SLA",
            "evidence": "Response time SLA metrics, Champion availability data",
            "category": "Champion Performance"
        },
        {
            "id": "2.10",
            "requirement": "Security Champion recommendations are tracked to implementation",
            "evidence": "Recommendation tracking system, implementation verification",
            "category": "Champion Follow-up"
        },
        {
            "id": "2.11",
            "requirement": "Clear escalation path exists from Champions to AppSec team",
            "evidence": "Escalation procedures, escalation records, contact information",
            "category": "Champion Process"
        },
        {
            "id": "2.12",
            "requirement": "Security Champions have authority to block releases for security issues",
            "evidence": "Authority documentation, examples of Champion-blocked releases",
            "category": "Champion Authority"
        },
        {
            "id": "2.13",
            "requirement": "Security Champion program effectiveness is measured with KPIs",
            "evidence": "Program metrics dashboard, effectiveness assessment reports",
            "category": "Champion Metrics"
        },
        {
            "id": "2.14",
            "requirement": "Regular Security Champion sync meetings are held",
            "evidence": "Meeting schedules, attendance records, meeting notes",
            "category": "Champion Program"
        },
        {
            "id": "2.15",
            "requirement": "Security Champion knowledge base and playbooks are maintained",
            "evidence": "Internal wiki, Security Champion playbooks, decision trees",
            "category": "Champion Resources"
        },
        {
            "id": "2.16",
            "requirement": "New Security Champions are onboarded with systematic training plan",
            "evidence": "Champion onboarding checklist, training plan documentation",
            "category": "Champion Training"
        },
        {
            "id": "2.17",
            "requirement": "Security Champion contributions are recognized and rewarded",
            "evidence": "Recognition program documentation, awards, acknowledgments",
            "category": "Champion Program"
        },
        {
            "id": "2.18",
            "requirement": "Security Champions receive regular updates on emerging threats",
            "evidence": "Threat intelligence briefings, security newsletters, training sessions",
            "category": "Champion Training"
        },
    ]


def get_domain3_requirements():
    """
    Domain 3: Security Testing - Unit & Integration
    
    Security tests at the unit and integration level catch issues early.
    But only if they actually test security-relevant behavior.
    """
    return [
        {
            "id": "3.1",
            "requirement": "Security test cases exist for all critical application functionality",
            "evidence": "Security test suite, test case repository, test documentation",
            "category": "Test Coverage"
        },
        {
            "id": "3.2",
            "requirement": "Input validation is tested at the unit test level",
            "evidence": "Input validation unit tests, boundary testing, malicious input tests",
            "category": "Unit Testing"
        },
        {
            "id": "3.3",
            "requirement": "Authentication logic is thoroughly tested with unit tests",
            "evidence": "Authentication unit test suites, auth failure test cases",
            "category": "Unit Testing"
        },
        {
            "id": "3.4",
            "requirement": "Authorization checks are tested at unit and integration levels",
            "evidence": "Authorization test cases, access control test matrices",
            "category": "Unit Testing"
        },
        {
            "id": "3.5",
            "requirement": "Session management security is tested",
            "evidence": "Session security test cases, timeout testing, session fixation tests",
            "category": "Integration Testing"
        },
        {
            "id": "3.6",
            "requirement": "Cryptographic functions are tested for correct implementation",
            "evidence": "Cryptography unit tests, key generation tests, encryption/decryption tests",
            "category": "Unit Testing"
        },
        {
            "id": "3.7",
            "requirement": "Error handling is tested to prevent information disclosure",
            "evidence": "Error handling test scenarios, exception testing, stack trace prevention tests",
            "category": "Unit Testing"
        },
        {
            "id": "3.8",
            "requirement": "Integration tests validate end-to-end authentication/authorization workflows",
            "evidence": "Integration test suites, workflow test scenarios",
            "category": "Integration Testing"
        },
        {
            "id": "3.9",
            "requirement": "Security test coverage is measured and tracked",
            "evidence": "Test coverage metrics for security-critical code, coverage reports",
            "category": "Test Metrics"
        },
        {
            "id": "3.10",
            "requirement": "Failed security tests block CI/CD pipeline deployment",
            "evidence": "Pipeline configuration, test failure blocking evidence, failed build logs",
            "category": "Test Enforcement"
        },
        {
            "id": "3.11",
            "requirement": "Security tests run automatically on every code commit",
            "evidence": "CI/CD configuration, test execution logs, automated test runs",
            "category": "Test Automation"
        },
        {
            "id": "3.12",
            "requirement": "Security test maintenance process exists and tests are kept current",
            "evidence": "Test maintenance procedures, test update logs, deprecation handling",
            "category": "Test Maintenance"
        },
        {
            "id": "3.13",
            "requirement": "Security test cases are peer-reviewed like production code",
            "evidence": "Test code review records, test PR approvals",
            "category": "Test Quality"
        },
        {
            "id": "3.14",
            "requirement": "Negative test cases are included for security-critical functions",
            "evidence": "Negative testing examples, failure mode testing",
            "category": "Unit Testing"
        },
        {
            "id": "3.15",
            "requirement": "Boundary condition testing is performed for input validation",
            "evidence": "Boundary test cases, edge case testing, overflow testing",
            "category": "Unit Testing"
        },
        {
            "id": "3.16",
            "requirement": "Race condition testing is performed where applicable",
            "evidence": "Concurrency test cases, threading tests, TOCTOU testing",
            "category": "Integration Testing"
        },
        {
            "id": "3.17",
            "requirement": "Security test results are archived for audit and trend analysis",
            "evidence": "Test result storage, historical test data, trend reports",
            "category": "Test Metrics"
        },
        {
            "id": "3.18",
            "requirement": "Test environment security is validated (no production data)",
            "evidence": "Test environment security assessment, data handling procedures",
            "category": "Test Environment"
        },
    ]


def get_domain4_requirements():
    """
    Domain 4: API & Application Security Testing
    
    APIs are attack surfaces. They need thorough security testing including
    authentication, authorization, input validation, and abuse scenarios.
    """
    return [
        {
            "id": "4.1",
            "requirement": "All APIs (REST, GraphQL, gRPC) have dedicated security test suites",
            "evidence": "API security test documentation, Postman/REST Assured collections",
            "category": "API Testing"
        },
        {
            "id": "4.2",
            "requirement": "API authentication mechanisms are thoroughly tested",
            "evidence": "Auth testing scenarios (OAuth, JWT, API keys), token validation tests",
            "category": "API Testing"
        },
        {
            "id": "4.3",
            "requirement": "Authorization is tested for all API endpoints and operations",
            "evidence": "Endpoint authorization test matrix, RBAC testing, privilege escalation tests",
            "category": "API Testing"
        },
        {
            "id": "4.4",
            "requirement": "Input validation is tested for all API parameters",
            "evidence": "Input validation test cases, parameter fuzzing, type confusion tests",
            "category": "API Testing"
        },
        {
            "id": "4.5",
            "requirement": "SQL injection testing is performed on all database-backed APIs",
            "evidence": "SQL injection test results, parameterized query validation",
            "category": "Injection Testing"
        },
        {
            "id": "4.6",
            "requirement": "Cross-Site Scripting (XSS) testing is performed",
            "evidence": "XSS test scenarios, output encoding validation, CSP testing",
            "category": "Injection Testing"
        },
        {
            "id": "4.7",
            "requirement": "XXE and deserialization vulnerability testing is performed",
            "evidence": "XXE test cases, unsafe deserialization tests, XML/JSON validation",
            "category": "Injection Testing"
        },
        {
            "id": "4.8",
            "requirement": "API fuzzing is performed to discover unexpected behavior",
            "evidence": "Fuzzing tool results (Burp Intruder, ffuf, RESTler), crash reports",
            "category": "API Testing"
        },
        {
            "id": "4.9",
            "requirement": "API rate limiting and abuse prevention are tested",
            "evidence": "Rate limit test evidence, DoS protection testing, throttling validation",
            "category": "API Testing"
        },
        {
            "id": "4.10",
            "requirement": "CORS configuration is validated for security",
            "evidence": "CORS test results, origin validation tests, preflight testing",
            "category": "API Testing"
        },
        {
            "id": "4.11",
            "requirement": "Security headers are validated in API responses",
            "evidence": "Header validation tests (HSTS, CSP, X-Frame-Options, etc.)",
            "category": "Application Testing"
        },
        {
            "id": "4.12",
            "requirement": "API error responses are tested for information disclosure",
            "evidence": "Error response security tests, stack trace prevention validation",
            "category": "API Testing"
        },
        {
            "id": "4.13",
            "requirement": "Staging/testing environment mirrors production security configuration",
            "evidence": "Environment parity documentation, configuration comparison",
            "category": "Test Environment"
        },
        {
            "id": "4.14",
            "requirement": "Test data is sanitized and does not include production data",
            "evidence": "Test data management policy, data anonymization procedures",
            "category": "Test Environment"
        },
        {
            "id": "4.15",
            "requirement": "API documentation is security-reviewed for sensitive information disclosure",
            "evidence": "API documentation review records, Swagger/OpenAPI security review",
            "category": "API Testing"
        },
        {
            "id": "4.16",
            "requirement": "GraphQL-specific security testing is performed (if applicable)",
            "evidence": "GraphQL introspection testing, query depth limiting, batching abuse tests",
            "category": "API Testing"
        },
        {
            "id": "4.17",
            "requirement": "WebSocket security is tested (if applicable)",
            "evidence": "WebSocket test scenarios, connection hijacking tests, message validation",
            "category": "Application Testing"
        },
        {
            "id": "4.18",
            "requirement": "API versioning security is validated across versions",
            "evidence": "Version security tests, deprecated endpoint validation, migration testing",
            "category": "API Testing"
        },
    ]


def get_domain5_requirements():
    """
    Domain 5: External Testing & Continuous Validation
    
    External perspectives find issues internal testing misses. Pentests and
    bug bounty programs provide reality checks on security posture.
    """
    return [
        {
            "id": "5.1",
            "requirement": "Annual penetration testing is performed for critical applications",
            "evidence": "Penetration test reports from last 12 months, test schedules",
            "category": "Penetration Testing"
        },
        {
            "id": "5.2",
            "requirement": "Penetration test scope covers all critical applications and APIs",
            "evidence": "Pentest scope documentation, application inventory",
            "category": "Penetration Testing"
        },
        {
            "id": "5.3",
            "requirement": "Penetration test findings are prioritized by risk and severity",
            "evidence": "Risk-ranked findings list, severity classification methodology",
            "category": "Penetration Testing"
        },
        {
            "id": "5.4",
            "requirement": "Critical pentest findings are remediated within 30 days",
            "evidence": "Remediation tracking system, SLA compliance metrics",
            "category": "Remediation"
        },
        {
            "id": "5.5",
            "requirement": "Penetration test retest validates that fixes are effective",
            "evidence": "Retest reports, fix validation evidence",
            "category": "Penetration Testing"
        },
        {
            "id": "5.6",
            "requirement": "Bug bounty or responsible disclosure program is active",
            "evidence": "Bug bounty platform documentation (HackerOne, Bugcrowd), program page",
            "category": "Bug Bounty"
        },
        {
            "id": "5.7",
            "requirement": "Vulnerability disclosure policy is publicly published",
            "evidence": "Published security.txt or vulnerability disclosure policy URL",
            "category": "Bug Bounty"
        },
        {
            "id": "5.8",
            "requirement": "External vulnerability submissions are triaged within 48 hours",
            "evidence": "Triage SLA metrics, response time tracking",
            "category": "Bug Bounty"
        },
        {
            "id": "5.9",
            "requirement": "Security researcher communication is handled professionally",
            "evidence": "Communication templates, researcher feedback, response examples",
            "category": "Bug Bounty"
        },
        {
            "id": "5.10",
            "requirement": "Bug bounty payments are tracked and processed timely",
            "evidence": "Payment records, payout tracking, bounty statistics",
            "category": "Bug Bounty"
        },
        {
            "id": "5.11",
            "requirement": "Security regression tests exist for all past critical vulnerabilities",
            "evidence": "Regression test suite, vulnerability-specific tests",
            "category": "Regression Testing"
        },
        {
            "id": "5.12",
            "requirement": "Security regression tests run automatically in CI/CD",
            "evidence": "CI/CD regression test execution logs, test automation",
            "category": "Regression Testing"
        },
        {
            "id": "5.13",
            "requirement": "Security testing metrics are tracked over time",
            "evidence": "Historical metrics dashboard, trend analysis",
            "category": "Metrics"
        },
        {
            "id": "5.14",
            "requirement": "Vulnerability trends are analyzed to identify patterns",
            "evidence": "Vulnerability trend analysis reports, root cause analysis",
            "category": "Metrics"
        },
        {
            "id": "5.15",
            "requirement": "Testing effectiveness is measured with KPIs",
            "evidence": "Testing effectiveness KPIs, detection rate metrics",
            "category": "Metrics"
        },
        {
            "id": "5.16",
            "requirement": "False negative analysis is performed after security incidents",
            "evidence": "Post-incident analysis reports, testing gap assessments",
            "category": "Continuous Improvement"
        },
        {
            "id": "5.17",
            "requirement": "Testing gaps identified in incidents are addressed systematically",
            "evidence": "Gap remediation plans, testing enhancement initiatives",
            "category": "Continuous Improvement"
        },
        {
            "id": "5.18",
            "requirement": "Continuous improvement process drives testing program evolution",
            "evidence": "Improvement backlog, testing maturity roadmap",
            "category": "Continuous Improvement"
        },
    ]

# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET CREATION
# ============================================================================

def create_instructions_sheet(ws, styles):
    """
    Create comprehensive Instructions sheet for Code Review & Testing assessment.
    
    Reviews and tests are the last lines of defense. This assessment evaluates
    whether those defenses actually work, not just whether they exist.
    """
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ISMS-IMP-A.8.28.3 – Code Review & Testing Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "ISO/IEC 27001:2022 - Control A.8.28: Secure Coding (Code Review & Testing)"
    apply_style(ws["A2"], styles["subheader"])
    ws.row_dimensions[2].height = 20

    # Document Information Block
    row = 4
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f"A{row}:F{row}")
    
    doc_info = [
        ("Document ID:", "ISMS-IMP-A.8.28.3"),
        ("Assessment Area:", "Secure Code Review & Security Testing"),
        ("Related Policy:", "ISMS-POL-A.8.28-S2.3 (Code Review & Testing)"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT - Enter assessment date]"),
        ("Completed By:", "[USER INPUT - Enter assessor name]"),
        ("Organization:", "[USER INPUT - Enter organization name]"),
        ("Review Cycle:", "Quarterly or when review/testing processes change"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"B{row}"] = value
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].font = Font(italic=True, size=10)
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    # How to Use This Workbook
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25

    instructions = [
        "1. Complete each assessment domain: Review Process → Champion Reviews → Unit/Integration Testing → API Testing → External Testing",
        "2. Focus on EFFECTIVENESS, not just process existence",
        "3. Ask: 'Does this review/test actually catch vulnerabilities before production?'",
        "4. Sample actual code reviews to validate quality (not just process documentation)",
        "5. Review test cases to ensure they test security-relevant behavior",
        "6. Analyze metrics: What % of vulnerabilities are caught in review vs. testing vs. production?",
        "7. Document gaps with specific, actionable remediation plans",
        "8. Obtain approvals from both Security and Engineering leadership",
        "",
        "\u26A0\uFE0F  CRITICAL: Avoid 'Review Theater' and 'Test Theater'",
        "   \u2022 'LGTM' without security analysis is not a review",
        "   \u2022 Tests that always pass don't test security",
        "   \u2022 Process without effectiveness is just cargo cult",
    ]

    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{row}:F{row}")
        if instruction.startswith("\u26A0\uFE0F"):
            ws[f"A{row}"].font = Font(bold=True, size=10, color="C00000")
        ws.row_dimensions[row].height = 30 if len(instruction) > 100 else 20
        row += 1

    # Assessment Domains Overview
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ASSESSMENT DOMAINS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Domain"
    ws[f"B{row}"] = "Focus Area"
    ws[f"C{row}"] = "Requirements"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    domain_overview = [
        ("1. Code Review Process & Governance", "Mandatory review, qualified reviewers, security checklists", "18"),
        ("2. Security Champion Review", "Champion program, architecture reviews, threat modeling", "18"),
        ("3. Security Testing - Unit & Integration", "Security test cases, auth/authz testing, test coverage", "18"),
        ("4. API & Application Security Testing", "API testing, injection testing, fuzzing, staging tests", "18"),
        ("5. External Testing & Continuous Validation", "Pentesting, bug bounty, regression tests, metrics", "18"),
    ]

    row += 1
    for domain, focus, req_count in domain_overview:
        ws[f"A{row}"] = domain
        ws[f"B{row}"] = focus
        ws[f"C{row}"] = req_count
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    ws[f"C{row}"] = "90 Total"
    ws[f"C{row}"].font = Font(bold=True)
    ws[f"C{row}"].alignment = Alignment(horizontal="center")

    # Implementation Status Legend
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "IMPLEMENTATION STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    legend_headers = ["Status", "Symbol", "Description", "When to Use"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    legend_data = [
        ("Implemented", "\u2705", "Review/test is effective at catching vulnerabilities", "When process exists AND demonstrably catches security issues"),
        ("Partially Implemented", "\u26A0\uFE0F", "Process exists but limited effectiveness", "When process exists but gaps in coverage or quality"),
        ("Not Implemented", "\u274C", "Review/test process does not exist", "When no process exists or it's completely ineffective"),
        ("N/A", "N/A", "Not applicable to environment", "Use sparingly - most review/testing is applicable"),
    ]

    row += 1
    for status, symbol, desc, when_use in legend_data:
        ws[f"A{row}"] = status
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = symbol
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = when_use
        
        if status == "Implemented":
            ws[f"A{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "Partially Implemented":
            ws[f"A{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif status == "Not Implemented":
            ws[f"A{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif status == "N/A":
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws.merge_cells(f"C{row}:F{row}")
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    # Acceptable Evidence Examples
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (Examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    evidence_examples = [
        "✓ Code review policy and procedures documentation",
        "✓ Sample code reviews showing security analysis (not just 'LGTM')",
        "✓ Repository settings enforcing review requirements",
        "✓ Security Champion roster and training records",
        "✓ Security-focused code review checklists (with usage evidence)",
        "✓ Security test suite code and test case documentation",
        "✓ Test execution logs showing security tests running",
        "✓ Test coverage reports (specifically for security tests)",
        "✓ API security test scenarios and results",
        "✓ Penetration test reports (last 12 months)",
        "✓ Bug bounty program documentation and submission records",
        "✓ Security regression test suite",
        "✓ Security metrics: vulnerabilities caught in review vs. testing vs. production",
        "✓ False negative analysis (post-incident reviews)",
        "✓ Remediation tracking for security findings",
    ]

    row += 1
    for evidence in evidence_examples:
        ws[f"A{row}"] = evidence
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

    # Common Anti-Patterns
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMMON ANTI-PATTERNS TO AVOID (Theater vs. Reality)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    antipatterns = [
        "\u274C REVIEW THEATER: 'LGTM' comments with no actual security analysis",
        "\u274C TEST THEATER: Tests that always pass because they don't test security",
        "\u274C Marking 'Implemented' because review policy exists (but nobody follows it)",
        "\u274C Security tests that are disabled 'temporarily' for months",
        "\u274C Pentests with artificially narrow scope to avoid finding issues",
        "\u274C Bug bounty program with no budget and no responses",
        "\u274C Reviewers who aren't trained on secure code review",
        "\u274C Measuring 'reviews performed' instead of 'vulnerabilities caught'",
        "\u274C Security Champions who are never consulted",
        "\u274C Test coverage metrics that don't include security test coverage",
    ]

    row += 1
    for antipattern in antipatterns:
        ws[f"A{row}"] = antipattern
        ws[f"A{row}"].font = Font(size=10, color="C00000")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Effectiveness Assessment Questions
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "KEY EFFECTIVENESS QUESTIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    effectiveness_questions = [
        "❓ What % of security issues are caught in code review?",
        "❓ What % of security issues are caught in testing?",
        "❓ What % of security issues reach production?",
        "❓ Are pentests finding issues internal testing should have caught?",
        "❓ Are bug bounty researchers finding trivial issues?",
        "❓ How many security regressions occur (previously fixed issues reintroduced)?",
        "❓ Do security tests actually test security-relevant behavior?",
        "❓ When was the last time a security test caught a real vulnerability?",
    ]

    row += 1
    for question in effectiveness_questions:
        ws[f"A{row}"] = question
        ws[f"A{row}"].font = Font(size=10, italic=True, color="2F5496")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 20
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 5: DOMAIN SHEET AND SUPPORTING SHEETS CREATION
# ============================================================================

def create_domain_sheet(ws, domain_name, requirements, styles):
    """Create standardized assessment domain sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Code Review & Testing Assessment: {domain_name}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    headers = ["ID", "Requirement", "Implementation Status", "Evidence Reference", "Comments", "Compliance"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        apply_style(cell, styles["column_header"])
    
    ws.row_dimensions[2].height = 30
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 12

    current_category = None
    row = 3
    
    for req in requirements:
        if "category" in req and req["category"] != current_category:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"▶ {req['category']}"
            ws[f"A{row}"].font = Font(bold=True, size=11, color="4472C4")
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20
            row += 1
            current_category = req["category"]
        
        ws[f"A{row}"] = req["id"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        ws[f"B{row}"] = req["requirement"]
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        validations["implementation_status"].add(ws[f"C{row}"])
        
        ws[f"D{row}"] = req.get("evidence", "")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"D{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[f"D{row}"].font = Font(size=9, italic=True, color="666666")
        
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"E{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws[f"F{row}"] = f'=IF(C{row}="Implemented","\u2705",IF(C{row}="Partially Implemented","\u26A0\uFE0F",IF(C{row}="Not Implemented","\u274C",IF(C{row}="N/A","N/A",""))))'
        ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws.row_dimensions[row].height = 45
        row += 1

    ws.freeze_panes = "A3"
    return ws


def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with compliance metrics."""
    
    ws.merge_cells("A1:E1")
    ws["A1"] = "Code Review & Testing Assessment - Summary Dashboard"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Date:"
    ws["B3"] = "=Instructions!B8"
    ws["A4"] = "Assessor:"
    ws["B4"] = "=Instructions!B9"
    ws["A5"] = "Organization:"
    ws["B5"] = "=Instructions!B10"
    
    for r in [3, 4, 5]:
        ws[f"A{r}"].font = Font(bold=True)

    ws["A7"] = "Overall Code Review & Testing Compliance:"
    ws["A7"].font = Font(bold=True, size=13)
    ws.merge_cells("A7:B7")
    
    ws["C7"] = """=(COUNTIF(Code_Review_Process!C:C,"Implemented")+
COUNTIF(Security_Champion_Review!C:C,"Implemented")+
COUNTIF(Unit_Integration_Testing!C:C,"Implemented")+
COUNTIF(API_Application_Testing!C:C,"Implemented")+
COUNTIF(External_Testing_Validation!C:C,"Implemented"))/
(COUNTA(Code_Review_Process!C:C)+
COUNTA(Security_Champion_Review!C:C)+
COUNTA(Unit_Integration_Testing!C:C)+
COUNTA(API_Application_Testing!C:C)+
COUNTA(External_Testing_Validation!C:C)-5-
COUNTIF(Code_Review_Process!C:C,"N/A")-
COUNTIF(Security_Champion_Review!C:C,"N/A")-
COUNTIF(Unit_Integration_Testing!C:C,"N/A")-
COUNTIF(API_Application_Testing!C:C,"N/A")-
COUNTIF(External_Testing_Validation!C:C,"N/A"))"""
    
    ws["C7"].number_format = "0%"
    ws["C7"].font = Font(bold=True, size=14)
    ws["C7"].alignment = Alignment(horizontal="center")
    
    ws["D7"] = '=IF(C7>=0.8,"🟢 Compliant",IF(C7>=0.6,"🟡 Needs Improvement","🔴 Non-Compliant"))'
    ws["D7"].font = Font(bold=True, size=12)
    ws["D7"].alignment = Alignment(horizontal="center")
    ws.merge_cells("D7:E7")

    ws["A10"] = "Domain Compliance Breakdown"
    ws["A10"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A10"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells("A10:E10")
    ws["A10"].alignment = Alignment(horizontal="center")

    domain_headers = ["Domain", "Compliance %", "Status", "Not Implemented", "Partially Implemented"]
    for col_num, header in enumerate(domain_headers, 1):
        cell = ws.cell(row=11, column=col_num, value=header)
        apply_style(cell, styles["column_header"])

    domains = [
        ("Code_Review_Process", "1. Code Review Process & Governance"),
        ("Security_Champion_Review", "2. Security Champion Review & Architecture"),
        ("Unit_Integration_Testing", "3. Security Testing - Unit & Integration"),
        ("API_Application_Testing", "4. API & Application Security Testing"),
        ("External_Testing_Validation", "5. External Testing & Continuous Validation"),
    ]

    row = 12
    for sheet_name, display_name in domains:
        ws[f"A{row}"] = display_name
        ws[f"B{row}"] = f'=COUNTIF({sheet_name}!C:C,"Implemented")/(COUNTA({sheet_name}!C:C)-1-COUNTIF({sheet_name}!C:C,"N/A"))'
        ws[f"B{row}"].number_format = "0%"
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"] = f'=IF(B{row}>=0.8,"🟢",IF(B{row}>=0.6,"🟡","🔴"))'
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"] = f'=COUNTIF({sheet_name}!C:C,"Not Implemented")'
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"E{row}"] = f'=COUNTIF({sheet_name}!C:C,"Partially Implemented")'
        ws[f"E{row}"].alignment = Alignment(horizontal="center")
        row += 1

    # Key Insights Section
    ws["A19"] = "Key Assessment Insights"
    ws["A19"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A19"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells("A19:E19")
    ws["A19"].alignment = Alignment(horizontal="center")

    insights = [
        "Focus on EFFECTIVENESS: Do reviews and tests actually catch vulnerabilities?",
        "Sample actual code reviews to validate quality of security analysis",
        "Review test cases to ensure they test security-relevant behavior",
        "Analyze: Where are vulnerabilities found? Review → Testing → Production?",
        "If pentests find trivial issues, internal testing is failing",
    ]

    row = 20
    for insight in insights:
        ws[f"A{row}"] = f"\u2022 {insight}"
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws[f"A{row}"].font = Font(italic=True, size=10)
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 25

    return ws


def create_evidence_register(ws, styles):
    """Create Evidence Register."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:G1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 25

    ws["A2"] = "Document all evidence supporting your Code Review & Testing assessment."
    ws.merge_cells("A2:G2")
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True)

    headers = ["Evidence ID", "Related Requirement", "Evidence Type", "Description", "Location/Link", "Collection Date", "Collected By"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        apply_style(cell, styles["column_header"])

    widths = {"A": 12, "B": 25, "C": 20, "D": 40, "E": 35, "F": 15, "G": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    validations["evidence_type"].add("C4:C200")
    ws.freeze_panes = "A4"
    return ws


def create_gap_analysis_sheet(ws, styles):
    """Create Gap Analysis sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:J1")
    ws["A1"] = "Gap Analysis & Remediation Plan"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 25

    headers = ["Gap ID", "Domain", "Requirement ID", "Requirement Description", "Current State", "Target State", "Priority", "Owner", "Target Date", "Status"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        apply_style(cell, styles["column_header"])

    widths = {"A": 10, "B": 22, "C": 12, "D": 35, "E": 25, "F": 25, "G": 12, "H": 20, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    validations["priority"].add("G4:G200")
    validations["gap_status"].add("J4:J200")
    ws.freeze_panes = "A4"
    return ws


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet."""
    
    ws.merge_cells("A1:D1")
    ws["A1"] = "Code Review & Testing Assessment - Approval Sign-Off"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws["A6"] = "Assessment Date:"
    ws["B6"] = "=Instructions!B8"
    ws["A7"] = "Assessor Name:"
    ws["B7"] = "=Instructions!B9"
    ws["A8"] = "Overall Compliance Score:"
    ws["B8"] = "=Summary_Dashboard!C7"
    ws["B8"].number_format = "0%"
    ws["B8"].font = Font(bold=True, size=12)

    for r in [6, 7, 8]:
        ws[f"A{r}"].font = Font(bold=True)

    ws["A11"] = "APPROVAL WORKFLOW"
    ws["A11"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A11"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells("A11:D11")
    ws["A11"].alignment = Alignment(horizontal="center")

    approval_headers = ["Role", "Name", "Signature", "Date"]
    for col_num, header in enumerate(approval_headers, 1):
        cell = ws.cell(row=12, column=col_num, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    approval_roles = ["Assessment Completer", "Application Security Lead", "Development Manager", "QA Manager", "CISO"]

    row = 13
    for role in approval_roles:
        ws[f"A{row}"] = role
        ws[f"A{row}"].font = Font(bold=True, size=10)
        for col in ["B", "C", "D"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15

    return ws

# ============================================================================
# SECTION 6: MAIN EXECUTION
# ============================================================================

def main():
    """
    Generate Code Review & Testing Assessment workbook.
    
    Reviews and tests are the last lines of defense before production.
    This assessment helps determine if those defenses actually work.
    """
    logger.info("=" * 80)
    logger.info(" " * 12 + "ISMS Control 8.28.3 - Code Review & Testing Assessment Generator")
    logger.info("=" * 80)
    logger.info("")
    
    try:
        logger.info("📝 [1/10] Creating workbook structure...")
        wb = create_workbook()
        styles = setup_styles()
        logger.info("     ✓ Workbook initialized with 10 sheets")

        logger.info("📄 [2/10] Creating Instructions sheet...")
        ws_instructions = wb["Instructions"]
        create_instructions_sheet(ws_instructions, styles)
        logger.info("     ✓ Instructions complete (focus: effectiveness, not theater)")

        logger.info("📊 [3/10] Creating Domain 1: Code Review Process & Governance...")
        ws_domain1 = wb["Code_Review_Process"]
        create_domain_sheet(ws_domain1, "Code Review Process & Governance", get_domain1_requirements(), styles)
        logger.info("     ✓ 18 requirements (Review Policy, Quality, Metrics)")

        logger.info("📊 [4/10] Creating Domain 2: Security Champion Review & Architecture...")
        ws_domain2 = wb["Security_Champion_Review"]
        create_domain_sheet(ws_domain2, "Security Champion Review & Architecture", get_domain2_requirements(), styles)
        logger.info("     ✓ 18 requirements (Champion Program, Threat Modeling, Design Review)")

        logger.info("📊 [5/10] Creating Domain 3: Security Testing - Unit & Integration...")
        ws_domain3 = wb["Unit_Integration_Testing"]
        create_domain_sheet(ws_domain3, "Security Testing - Unit & Integration", get_domain3_requirements(), styles)
        logger.info("     ✓ 18 requirements (Unit Tests, Integration Tests, Test Coverage)")

        logger.info("📊 [6/10] Creating Domain 4: API & Application Security Testing...")
        ws_domain4 = wb["API_Application_Testing"]
        create_domain_sheet(ws_domain4, "API & Application Security Testing", get_domain4_requirements(), styles)
        logger.info("     ✓ 18 requirements (API Testing, Injection Testing, Fuzzing)")

        logger.info("📊 [7/10] Creating Domain 5: External Testing & Continuous Validation...")
        ws_domain5 = wb["External_Testing_Validation"]
        create_domain_sheet(ws_domain5, "External Testing & Continuous Validation", get_domain5_requirements(), styles)
        logger.info("     ✓ 18 requirements (Pentesting, Bug Bounty, Regression Tests)")

        logger.info("📈 [8/10] Creating Summary Dashboard...")
        ws_summary = wb["Summary_Dashboard"]
        create_summary_dashboard(ws_summary, styles)
        logger.info("     ✓ Executive summary with effectiveness insights")

        logger.info("📎 [9/10] Creating Evidence Register...")
        ws_evidence = wb["Evidence_Register"]
        create_evidence_register(ws_evidence, styles)
        logger.info("     ✓ Evidence tracking")

        logger.info("\u26A0\uFE0F  [10/10] Creating Gap Analysis & Approval sheets...")
        ws_gap = wb["Gap_Analysis"]
        create_gap_analysis_sheet(ws_gap, styles)
        
        ws_approval = wb["Approval_Sign_Off"]
        create_approval_sheet(ws_approval, styles)
        logger.info("     ✓ Gap tracking and approval workflow")

        logger.info("")
        logger.info("💾 Saving workbook...")
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.8.28.3_Code_Review_Testing_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("")
        logger.info("=" * 80)
        logger.info("\u2705 SUCCESS: Code Review & Testing Assessment workbook generated!")
        logger.info("=" * 80)
        logger.info("")
        logger.info(f"📁 File: {filename}")
        logger.info(f"📊 Total Requirements: 90 (18 per domain × 5 domains)")
        logger.info("")
        logger.info("📌 Assessment Domains:")
        logger.info("   1. Code Review Process & Governance    (18 requirements)")
        logger.info("   2. Security Champion Review            (18 requirements)")
        logger.info("   3. Unit & Integration Testing          (18 requirements)")
        logger.info("   4. API & Application Testing           (18 requirements)")
        logger.info("   5. External Testing & Validation       (18 requirements)")
        logger.info("")
        logger.info("💡 Key Focus: Does it catch vulnerabilities? Not just 'does the process exist?'")
        logger.info("")
        logger.info("\u26A0\uFE0F  Remember:")
        logger.info("   \u2022 'LGTM' without security analysis is not a review")
        logger.info("   \u2022 Tests that always pass don't test security")
        logger.info("   \u2022 Avoid Review Theater and Test Theater")
        logger.info("   \u2022 Focus on EFFECTIVENESS, not just process existence")
        logger.info("")
        logger.info("📊 Effectiveness Questions to Answer:")
        logger.info("   \u2022 What % of vulnerabilities are caught in review vs. testing vs. production?")
        logger.info("   \u2022 Are pentests finding issues internal testing should have caught?")
        logger.info("   \u2022 How many security regressions occur?")
        logger.info("")
        logger.info("=" * 80)
        
        return 0

    except Exception as e:
        logger.info("")
        logger.info("=" * 80)
        logger.error("\u274C ERROR: Failed to generate workbook")
        logger.info("=" * 80)
        logger.error(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
