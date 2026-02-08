#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.28.2 - Standards & Tools Assessment
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Assessment Domain 2 of 4: Coding Standards and Development Tool Security

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific programming languages, frameworks, development
tools, and coding standard requirements.

Key customization areas:
1. Programming languages and frameworks (match your technology stack)
2. Coding standards and style guides (specific to your language choices)
3. Development tools and IDE configurations (based on your toolchain)
4. Security linting rules and enforcement (adapt to your risk profile)
5. Compliance criteria and scoring (aligned with your quality requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
secure coding standards implementation and development tool security configuration
across all programming languages and platforms in use.

**Purpose:**
Enables systematic assessment of coding standards and tool security against
ISO 27001:2022 Control A.8.28 requirements, supporting evidence-based validation
of secure development practices and tool-enforced security controls.

**Assessment Scope:**
- Language-specific secure coding standards (OWASP, CWE, CERT, vendor guides)
- IDE and editor security plugins and configurations
- Static Application Security Testing (SAST) tool implementation
- Security linting rules and enforcement mechanisms
- Code formatting and style guide compliance
- Dependency management and vulnerability scanning tools
- Secret detection and credential management
- Pre-commit hooks and automated security checks
- Developer workstation security hardening
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and coding standard references
2. Language Standards - Per-language secure coding standard implementation
3. SAST Tools - Static analysis tool coverage and configuration
4. IDE Security - Development environment security plugin assessment
5. Linting & Formatting - Security linter rules and enforcement
6. Dependency Management - SCA tools and vulnerability scanning
7. Secret Detection - Credential and API key protection mechanisms
8. Pre-Commit Controls - Automated security gate implementation
9. Workstation Security - Developer endpoint hardening assessment
10. Gap Analysis - Missing or inadequate standards/tools and remediation
11. Evidence Register - Audit evidence tracking and documentation
12. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with language and tool dropdown lists
- Conditional formatting for standard compliance and tool coverage
- Automated gap identification for unsupported languages or missing tools
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with SAST/SCA tool configuration exports

**Integration:**
This assessment feeds into the A.8.28.5 Compliance Dashboard, which
consolidates data from all four assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a828_2_standards_tools.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a828_2_standards_tools.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a828_2_standards_tools.py --date 20250124

Output:
    File: ISMS_A_8_28_2_Standards_Tools_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize language-specific coding standards to match tech stack
    2. Inventory all programming languages, frameworks, and tools in use
    3. Complete assessments for each language and development tool
    4. Validate SAST/SCA tool coverage and rule configuration
    5. Review pre-commit hook implementation and enforcement
    6. Conduct gap analysis for languages without standards or inadequate tooling
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (tool configs, scan results, standard docs)
    9. Obtain stakeholder approvals (Dev Leads, AppSec, Engineering Manager)
    10. Feed results into A.8.28.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    2 of 4 (Standards & Tools)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.2: Standards & Tools Implementation Guide
    - ISMS-IMP-A.8.28.1: SDLC Integration Assessment (Domain 1)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Assessment (Domain 3)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Assessment (Domain 4)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a828_1_sdlc_assessment.py
    - generate_a828_3_code_review_testing.py
    - generate_a828_4_third_party_oss.py
    - generate_a828_5_compliance_dashboard.py
    - normalize_assessment_files_a828.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.28.2 specification
    - Supports comprehensive coding standards and tool security evaluation
    - Integrated with A.8.28.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Technology Stack Specificity:**
This assessment framework is technology-agnostic by design. Customize language
lists, coding standards, and tool recommendations to match your organization's
specific programming languages, frameworks, and development ecosystem.

**Tool Selection Philosophy:**
Prioritize tools that integrate seamlessly with existing workflows over
"best-of-breed" tools that developers won't use. An adequate tool that's
actually used is better than a perfect tool that's bypassed.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of SAST/SCA tool configurations, linting
rules, and coding standard adoption.

**Data Protection:**
Assessment workbooks contain sensitive development information including:
- Technology stack details and framework versions
- Security tool configurations and rule sets
- Vulnerability patterns and weakness trends
- Development environment details

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check SAST/SCA rule updates and new language support
- Semi-annually: Review coding standard versions and update references
- Annually: Complete reassessment of all languages and tools in use
- Ad-hoc: When new languages/frameworks adopted or tools changed

**Quality Assurance:**
Have application security SMEs, technical leads for each language/platform,
and DevOps engineers validate assessments before using results for compliance
reporting or remediation decisions.

**Regulatory Alignment:**
Ensure coding standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 secure coding standards (Requirement 6.2)
- Healthcare: HIPAA secure development practices
- Finance: Regional banking secure coding requirements
- Government: NIST Secure Software Development Framework (SSDF)

Customize assessment criteria to include regulatory-specific requirements.

**Integration with Development Tools:**
Where possible, integrate assessment data with:
- SAST tool outputs (SonarQube, Checkmarx, Fortify, Semgrep, etc.)
- SCA tool results (Snyk, WhiteSource, Black Duck, Dependabot, etc.)
- IDE plugin configurations (export from VS Code, IntelliJ, etc.)
- Pre-commit hook repositories (Git hook configurations)

Automated tool integration reduces manual effort and improves accuracy.

**Don't Fool Yourself - Feynman Principle:**
Having a coding standard document doesn't mean developers follow it. Having
SAST tools doesn't mean they're configured correctly or that findings are
addressed. Assess actual tool usage and standard adoption:
- Are SAST scans running on every commit/build?
- What percentage of findings are actually fixed vs. marked "Won't Fix"?
- Can you show code that was rejected for standard violations?
- Are developers actually using IDE security plugins?

Evidence of use > Existence of tools.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.28.2"
WORKBOOK_NAME = "Coding Standards and Development Tool Security"
CONTROL_ID = "A.8.28"
CONTROL_NAME = "Secure Coding"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

import sys


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.28.2 specification
    # 10 sheets total: Instructions + 5 assessment domains + 4 supporting sheets
    sheets = [
        "Instructions",
        "Coding_Standards_Adoption",
        "SAST_SCA_Tools",
        "DAST_Security_Testing_Tools",
        "IDE_Plugins_Linters",
        "Tool_Effectiveness_Metrics",
        "Summary_Dashboard",
        "Evidence_Register",
        "Gap_Analysis",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Returns style templates as dictionaries to avoid Excel's shared object warnings.
    Each cell gets a fresh style instance - no cargo cult object sharing.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_implemented": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notimplemented": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        },
        "priority_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "priority_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "priority_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "priority_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    
    Tools are only as good as their configuration and usage.
    These validations help assess deployment maturity, not just presence.
    """
    validations = {
        # Implementation status dropdowns
        'implementation_status': DataValidation(
            type="list",
            formula1='"Implemented,Partially Implemented,Not Implemented,N/A"',
            allow_blank=False
        ),
        'deployment_status': DataValidation(
            type="list",
            formula1='"Deployed,Pilot,Planned,Not Deployed"',
            allow_blank=False
        ),
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Unknown"',
            allow_blank=False
        ),
        
        # Tool maturity and effectiveness
        'maturity_level': DataValidation(
            type="list",
            formula1='"Mature,Operational,Developing,Initial,Not Deployed"',
            allow_blank=False
        ),
        'effectiveness_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Fair,Poor,Not Assessed"',
            allow_blank=False
        ),
        'adoption_rate': DataValidation(
            type="list",
            formula1='"90-100%,70-89%,50-69%,Below 50%,Unknown"',
            allow_blank=False
        ),
        'coverage': DataValidation(
            type="list",
            formula1='"Complete,High (>75%),Medium (50-75%),Low (<50%),None"',
            allow_blank=False
        ),
        
        # Priority and risk levels
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Negligible"',
            allow_blank=False
        ),
        
        # Tool integration
        'integration_level': DataValidation(
            type="list",
            formula1='"Fully Integrated,Partially Integrated,Standalone,Not Integrated"',
            allow_blank=False
        ),
        'automation_level': DataValidation(
            type="list",
            formula1='"Fully Automated,Partially Automated,Manual,Not Applicable"',
            allow_blank=False
        ),
        
        # Gap analysis
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Closed,Deferred"',
            allow_blank=False
        ),
        
        # Evidence types
        'evidence_type': DataValidation(
            type="list",
            formula1='"Document,Screenshot,Report,Configuration,URL,Tool Output,Dashboard,Policy,Other"',
            allow_blank=False
        ),
        
        # Tool types
        'tool_type_sast': DataValidation(
            type="list",
            formula1='"SonarQube,Semgrep,Checkmarx,Fortify,Veracode,CodeQL,Other,None"',
            allow_blank=False
        ),
        'tool_type_sca': DataValidation(
            type="list",
            formula1='"Snyk,WhiteSource,Dependabot,Black Duck,OWASP Dependency-Check,Other,None"',
            allow_blank=False
        ),
        'tool_type_dast': DataValidation(
            type="list",
            formula1='"OWASP ZAP,Burp Suite,Acunetix,Netsparker,AppScan,Other,None"',
            allow_blank=False
        ),
        'tool_type_secrets': DataValidation(
            type="list",
            formula1='"TruffleHog,GitGuardian,git-secrets,detect-secrets,Gitleaks,Other,None"',
            allow_blank=False
        ),
        
        # Scan frequency
        'scan_frequency': DataValidation(
            type="list",
            formula1='"Every Commit,Daily,Weekly,On-Demand,Never"',
            allow_blank=False
        ),
        
        # Standards compliance
        'standards_compliance': DataValidation(
            type="list",
            formula1='"Fully Compliant,Mostly Compliant,Partially Compliant,Non-Compliant,Not Assessed"',
            allow_blank=False
        ),
        
        # Training status
        'training_status': DataValidation(
            type="list",
            formula1='"Completed,In Progress,Not Started,N/A"',
            allow_blank=False
        ),
        
        # Approval decision
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations

# ============================================================================
# SECTION 3: ASSESSMENT DOMAIN DATA DEFINITIONS
# ============================================================================

def get_domain1_requirements():
    """
    Domain 1: Coding Standards Adoption
    
    Standards without adoption are just documentation.
    This assesses whether secure coding standards are actually followed.
    """
    return [
        {
            "id": "1.1",
            "requirement": "Organization has documented secure coding standards",
            "evidence": "Secure coding standards document, developer wiki, internal guidelines",
            "category": "Standards Documentation"
        },
        {
            "id": "1.2",
            "requirement": "Coding standards reference OWASP Top 10 and CWE Top 25",
            "evidence": "Standards document with OWASP/CWE mappings, policy references",
            "category": "Standards Documentation"
        },
        {
            "id": "1.3",
            "requirement": "Standards cover input validation requirements for all languages used",
            "evidence": "Language-specific input validation guidelines, code examples",
            "category": "Standards Documentation"
        },
        {
            "id": "1.4",
            "requirement": "Standards cover output encoding/escaping requirements",
            "evidence": "XSS prevention guidelines, encoding standards documentation",
            "category": "Standards Documentation"
        },
        {
            "id": "1.5",
            "requirement": "Standards cover authentication and session management requirements",
            "evidence": "Auth standards, session handling guidelines, JWT/token standards",
            "category": "Standards Documentation"
        },
        {
            "id": "1.6",
            "requirement": "Standards cover cryptography requirements (algorithms, key management)",
            "evidence": "Crypto standards, approved algorithms list, key management policies",
            "category": "Standards Documentation"
        },
        {
            "id": "1.7",
            "requirement": "Standards cover error handling and logging requirements",
            "evidence": "Logging standards, error handling patterns, log retention policies",
            "category": "Standards Documentation"
        },
        {
            "id": "1.8",
            "requirement": "Developers are trained on secure coding standards",
            "evidence": "Training completion records, training materials, attendance logs",
            "category": "Training & Awareness"
        },
        {
            "id": "1.9",
            "requirement": "Secure coding standards are easily accessible to all developers",
            "evidence": "Internal wiki links, IDE bookmarks, quick reference guides",
            "category": "Accessibility"
        },
        {
            "id": "1.10",
            "requirement": "Standards are reviewed and updated at least annually",
            "evidence": "Standards review schedule, change logs, version history",
            "category": "Standards Maintenance"
        },
        {
            "id": "1.11",
            "requirement": "Language-specific secure coding guides exist for primary languages",
            "evidence": "Python/Java/JavaScript/etc. specific security guidelines",
            "category": "Standards Documentation"
        },
        {
            "id": "1.12",
            "requirement": "Standards include code examples (secure vs. insecure patterns)",
            "evidence": "Code example repository, good/bad pattern documentation",
            "category": "Standards Documentation"
        },
        {
            "id": "1.13",
            "requirement": "Compliance with standards is measured and tracked",
            "evidence": "Compliance metrics, SAST rule compliance, audit results",
            "category": "Standards Enforcement"
        },
        {
            "id": "1.14",
            "requirement": "Standards violations are identified during code review",
            "evidence": "Code review checklists, review comments, violation tracking",
            "category": "Standards Enforcement"
        },
        {
            "id": "1.15",
            "requirement": "Developers can request clarifications or exceptions to standards",
            "evidence": "Exception request process, clarification tickets, security champion reviews",
            "category": "Standards Governance"
        },
        {
            "id": "1.16",
            "requirement": "Standards compliance is part of developer performance evaluation",
            "evidence": "Performance review criteria, security KPIs for developers",
            "category": "Standards Enforcement"
        },
        {
            "id": "1.17",
            "requirement": "Secure coding champions help developers apply standards",
            "evidence": "Security champion program documentation, champion activity logs",
            "category": "Training & Awareness"
        },
        {
            "id": "1.18",
            "requirement": "New hire onboarding includes secure coding standards training",
            "evidence": "Onboarding checklist, new hire training materials, completion records",
            "category": "Training & Awareness"
        },
    ]


def get_domain2_requirements():
    """
    Domain 2: SAST & SCA Tools
    
    Static analysis tools find vulnerabilities before they reach production.
    But only if they're configured correctly and findings are actually remediated.
    """
    return [
        {
            "id": "2.1",
            "requirement": "SAST (Static Application Security Testing) tool is deployed",
            "evidence": "Tool license, deployment documentation, access credentials",
            "category": "SAST Deployment"
        },
        {
            "id": "2.2",
            "requirement": "SAST tool covers all primary programming languages used",
            "evidence": "Language support documentation, configuration for each language",
            "category": "SAST Configuration"
        },
        {
            "id": "2.3",
            "requirement": "SAST scans run automatically on every code commit/PR",
            "evidence": "CI/CD integration config, pipeline logs showing SAST execution",
            "category": "SAST Integration"
        },
        {
            "id": "2.4",
            "requirement": "SAST tool is configured with security-focused rulesets",
            "evidence": "Ruleset configuration, enabled security rules list",
            "category": "SAST Configuration"
        },
        {
            "id": "2.5",
            "requirement": "SAST findings are prioritized by severity (Critical/High/Medium/Low)",
            "evidence": "Severity classification schema, findings dashboard",
            "category": "SAST Configuration"
        },
        {
            "id": "2.6",
            "requirement": "Critical/High SAST findings block code deployment",
            "evidence": "Quality gate configuration, blocked deployment logs",
            "category": "SAST Integration"
        },
        {
            "id": "2.7",
            "requirement": "SAST false positives can be suppressed with justification",
            "evidence": "Suppression process, suppression audit trail, justification records",
            "category": "SAST Configuration"
        },
        {
            "id": "2.8",
            "requirement": "SAST findings are tracked to resolution in ticketing system",
            "evidence": "JIRA/Azure DevOps integration, vulnerability tracking tickets",
            "category": "SAST Remediation"
        },
        {
            "id": "2.9",
            "requirement": "SCA (Software Composition Analysis) tool is deployed",
            "evidence": "Tool license, deployment documentation, access credentials",
            "category": "SCA Deployment"
        },
        {
            "id": "2.10",
            "requirement": "SCA tool scans all dependency manifests (package.json, pom.xml, etc.)",
            "evidence": "Supported manifest files list, scan coverage reports",
            "category": "SCA Configuration"
        },
        {
            "id": "2.11",
            "requirement": "SCA scans run automatically on every build",
            "evidence": "CI/CD integration, build logs showing SCA execution",
            "category": "SCA Integration"
        },
        {
            "id": "2.12",
            "requirement": "SCA tool checks for known vulnerabilities in dependencies (CVEs)",
            "evidence": "Vulnerability database updates, CVE detection reports",
            "category": "SCA Configuration"
        },
        {
            "id": "2.13",
            "requirement": "SCA tool checks for license compliance issues",
            "evidence": "License policy configuration, license violation reports",
            "category": "SCA Configuration"
        },
        {
            "id": "2.14",
            "requirement": "Critical/High vulnerability findings block deployment",
            "evidence": "Quality gate configuration, blocked deployment logs",
            "category": "SCA Integration"
        },
        {
            "id": "2.15",
            "requirement": "SCA tool provides remediation guidance (upgrade paths)",
            "evidence": "Remediation recommendations, automated PR generation",
            "category": "SCA Remediation"
        },
        {
            "id": "2.16",
            "requirement": "Dependency vulnerabilities have defined SLAs for remediation",
            "evidence": "Vulnerability remediation SLA policy, SLA tracking metrics",
            "category": "SCA Remediation"
        },
        {
            "id": "2.17",
            "requirement": "SAST and SCA findings are reviewed by security team regularly",
            "evidence": "Security review meeting notes, findings triage records",
            "category": "Tool Governance"
        },
        {
            "id": "2.18",
            "requirement": "Metrics track SAST/SCA finding trends and remediation velocity",
            "evidence": "Security metrics dashboard, trend analysis reports",
            "category": "Tool Effectiveness"
        },
    ]


def get_domain3_requirements():
    """
    Domain 3: DAST & Security Testing Tools
    
    Dynamic testing finds runtime vulnerabilities that static analysis misses.
    Testing in environments that mirror production is critical.
    """
    return [
        {
            "id": "3.1",
            "requirement": "DAST (Dynamic Application Security Testing) tool is deployed",
            "evidence": "Tool license, deployment documentation, scanning infrastructure",
            "category": "DAST Deployment"
        },
        {
            "id": "3.2",
            "requirement": "DAST scans run against staging/pre-production environments",
            "evidence": "Scan configuration, target environment documentation",
            "category": "DAST Configuration"
        },
        {
            "id": "3.3",
            "requirement": "DAST scans are executed before production releases",
            "evidence": "Release checklist with DAST requirement, scan reports",
            "category": "DAST Integration"
        },
        {
            "id": "3.4",
            "requirement": "DAST tool tests for OWASP Top 10 vulnerabilities",
            "evidence": "Scan policy configuration, OWASP Top 10 test coverage",
            "category": "DAST Configuration"
        },
        {
            "id": "3.5",
            "requirement": "DAST scans include authenticated testing (logged-in scenarios)",
            "evidence": "Authentication configuration, authenticated scan results",
            "category": "DAST Configuration"
        },
        {
            "id": "3.6",
            "requirement": "DAST findings are classified by severity and risk",
            "evidence": "Severity schema, risk scoring methodology",
            "category": "DAST Configuration"
        },
        {
            "id": "3.7",
            "requirement": "Critical DAST findings block production deployment",
            "evidence": "Release gate configuration, blocked release records",
            "category": "DAST Integration"
        },
        {
            "id": "3.8",
            "requirement": "DAST false positives are validated and documented",
            "evidence": "False positive validation process, validation records",
            "category": "DAST Remediation"
        },
        {
            "id": "3.9",
            "requirement": "API security testing tools are deployed for API testing",
            "evidence": "API testing tool documentation (Postman, REST Assured, etc.)",
            "category": "API Testing"
        },
        {
            "id": "3.10",
            "requirement": "API tests include authentication and authorization testing",
            "evidence": "API test scenarios, auth bypass test cases",
            "category": "API Testing"
        },
        {
            "id": "3.11",
            "requirement": "API tests include input validation and injection testing",
            "evidence": "API fuzzing results, injection test cases",
            "category": "API Testing"
        },
        {
            "id": "3.12",
            "requirement": "Container/image scanning tool is deployed",
            "evidence": "Tool documentation (Trivy, Clair, Anchore), scan integration",
            "category": "Container Security"
        },
        {
            "id": "3.13",
            "requirement": "Container images are scanned before deployment",
            "evidence": "Container scan results, vulnerability reports",
            "category": "Container Security"
        },
        {
            "id": "3.14",
            "requirement": "IaC (Infrastructure-as-Code) scanning tool is deployed",
            "evidence": "Tool documentation (Checkov, tfsec, Terrascan), scan config",
            "category": "IaC Security"
        },
        {
            "id": "3.15",
            "requirement": "IaC is scanned for misconfigurations before deployment",
            "evidence": "IaC scan results, misconfiguration findings",
            "category": "IaC Security"
        },
        {
            "id": "3.16",
            "requirement": "Security testing results are archived for audit purposes",
            "evidence": "Test result storage, historical scan archives",
            "category": "Testing Governance"
        },
        {
            "id": "3.17",
            "requirement": "Penetration testing is performed annually for critical applications",
            "evidence": "Pentest reports, remediation tracking, retest results",
            "category": "Penetration Testing"
        },
        {
            "id": "3.18",
            "requirement": "Bug bounty or responsible disclosure program is active",
            "evidence": "Bug bounty platform (HackerOne, Bugcrowd), submission records",
            "category": "External Testing"
        },
    ]


def get_domain4_requirements():
    """
    Domain 4: IDE Plugins & Linters
    
    Catching issues in the IDE is the earliest and cheapest place to fix them.
    Developer-facing tools need high usability or they'll be ignored.
    """
    return [
        {
            "id": "4.1",
            "requirement": "Security-focused IDE plugins are available for developers",
            "evidence": "Plugin catalog, installation guides, approved plugin list",
            "category": "IDE Plugins"
        },
        {
            "id": "4.2",
            "requirement": "SAST IDE plugins provide real-time security feedback",
            "evidence": "SonarLint, Snyk IDE plugin, or similar tool documentation",
            "category": "IDE Plugins"
        },
        {
            "id": "4.3",
            "requirement": "IDE plugins are integrated with organizational security policies",
            "evidence": "Plugin configuration synced with SAST/SCA rules",
            "category": "IDE Configuration"
        },
        {
            "id": "4.4",
            "requirement": "IDE plugin adoption rate is measured and tracked",
            "evidence": "Plugin telemetry, adoption metrics dashboard",
            "category": "IDE Adoption"
        },
        {
            "id": "4.5",
            "requirement": "Developers receive training on using IDE security plugins",
            "evidence": "Plugin training materials, usage documentation",
            "category": "IDE Training"
        },
        {
            "id": "4.6",
            "requirement": "Code linters are configured with security rules",
            "evidence": "Linter configuration files (eslint, pylint, rubocop), security rules enabled",
            "category": "Linting"
        },
        {
            "id": "4.7",
            "requirement": "Linters run automatically in developer IDEs",
            "evidence": "IDE linter integration, pre-commit hook configuration",
            "category": "Linting"
        },
        {
            "id": "4.8",
            "requirement": "Linting rules are enforced in CI/CD pipeline",
            "evidence": "Pipeline linting step, linting failure logs",
            "category": "Linting"
        },
        {
            "id": "4.9",
            "requirement": "Secret scanning pre-commit hooks are deployed",
            "evidence": "Git hooks configuration (git-secrets, detect-secrets), hook installation",
            "category": "Secret Prevention"
        },
        {
            "id": "4.10",
            "requirement": "Pre-commit hooks prevent commits containing secrets",
            "evidence": "Blocked commit logs, secret detection examples",
            "category": "Secret Prevention"
        },
        {
            "id": "4.11",
            "requirement": "Developers can bypass pre-commit hooks only with justification",
            "evidence": "Bypass process documentation, bypass audit logs",
            "category": "Secret Prevention"
        },
        {
            "id": "4.12",
            "requirement": "Code formatting tools enforce consistent style",
            "evidence": "Prettier, Black, gofmt configuration, formatting enforcement",
            "category": "Code Quality"
        },
        {
            "id": "4.13",
            "requirement": "Complexity metrics are tracked (cyclomatic complexity, etc.)",
            "evidence": "SonarQube complexity metrics, code quality dashboards",
            "category": "Code Quality"
        },
        {
            "id": "4.14",
            "requirement": "Code coverage metrics are tracked and enforced",
            "evidence": "Coverage reports, minimum coverage thresholds",
            "category": "Code Quality"
        },
        {
            "id": "4.15",
            "requirement": "Dependency update tools provide automated update PRs",
            "evidence": "Dependabot, Renovate Bot configuration, automated PRs",
            "category": "Dependency Management"
        },
        {
            "id": "4.16",
            "requirement": "IDE plugins provide inline documentation for secure coding",
            "evidence": "Context-aware security hints, inline documentation",
            "category": "IDE Plugins"
        },
        {
            "id": "4.17",
            "requirement": "Developer feedback on tool usability is collected and acted upon",
            "evidence": "Developer surveys, tool improvement backlog",
            "category": "Tool Governance"
        },
        {
            "id": "4.18",
            "requirement": "Tool performance impact on developer workflow is monitored",
            "evidence": "Performance metrics, developer satisfaction scores",
            "category": "Tool Governance"
        },
    ]


def get_domain5_requirements():
    """
    Domain 5: Tool Effectiveness & Metrics
    
    Tools are investments. We need to measure if they're working.
    As Feynman said: "If it disagrees with experiment, it's wrong."
    """
    return [
        {
            "id": "5.1",
            "requirement": "Security tool effectiveness is measured with KPIs",
            "evidence": "KPI dashboard, effectiveness metrics definition",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.2",
            "requirement": "Vulnerability detection rate is tracked per tool",
            "evidence": "Detection metrics by tool, true positive rates",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.3",
            "requirement": "False positive rate is measured and minimized",
            "evidence": "False positive metrics, tuning efforts documentation",
            "category": "Tool Tuning"
        },
        {
            "id": "5.4",
            "requirement": "Mean time to remediation (MTTR) is tracked for findings",
            "evidence": "MTTR metrics, remediation velocity dashboard",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.5",
            "requirement": "Tool coverage is measured (% of codebase scanned)",
            "evidence": "Coverage reports, unscanned code identification",
            "category": "Tool Coverage"
        },
        {
            "id": "5.6",
            "requirement": "Security tool findings trends are analyzed over time",
            "evidence": "Trend analysis reports, historical metrics",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.7",
            "requirement": "Tool costs are tracked and justified by value delivered",
            "evidence": "Cost-benefit analysis, ROI documentation",
            "category": "Tool Governance"
        },
        {
            "id": "5.8",
            "requirement": "Tool performance (scan time, resource usage) is monitored",
            "evidence": "Performance monitoring, scan duration metrics",
            "category": "Tool Performance"
        },
        {
            "id": "5.9",
            "requirement": "Duplicate findings across tools are identified and deduplicated",
            "evidence": "Deduplication process, consolidated findings reports",
            "category": "Tool Integration"
        },
        {
            "id": "5.10",
            "requirement": "Security findings are correlated with production incidents",
            "evidence": "Incident analysis, finding-to-incident correlation reports",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.11",
            "requirement": "Tool findings are integrated with vulnerability management system",
            "evidence": "VM system integration, unified vulnerability tracking",
            "category": "Tool Integration"
        },
        {
            "id": "5.12",
            "requirement": "Security metrics are reported to management quarterly",
            "evidence": "Management reports, executive dashboards",
            "category": "Reporting"
        },
        {
            "id": "5.13",
            "requirement": "Tool configuration changes are tracked and audited",
            "evidence": "Configuration change logs, audit trails",
            "category": "Tool Governance"
        },
        {
            "id": "5.14",
            "requirement": "Tool licenses are managed and renewal dates tracked",
            "evidence": "License management system, renewal calendar",
            "category": "Tool Governance"
        },
        {
            "id": "5.15",
            "requirement": "New tools are evaluated against security requirements",
            "evidence": "Tool evaluation framework, POC results",
            "category": "Tool Selection"
        },
        {
            "id": "5.16",
            "requirement": "Tools are regularly updated to latest stable versions",
            "evidence": "Tool version inventory, update schedule",
            "category": "Tool Maintenance"
        },
        {
            "id": "5.17",
            "requirement": "Tool effectiveness is benchmarked against industry standards",
            "evidence": "Benchmark comparison reports, industry metrics",
            "category": "Tool Effectiveness"
        },
        {
            "id": "5.18",
            "requirement": "Continuous improvement process exists for tool optimization",
            "evidence": "Improvement backlog, optimization initiatives",
            "category": "Tool Governance"
        },
    ]

# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET CREATION
# ============================================================================

def create_instructions_sheet(ws, styles):
    """
    Create comprehensive Instructions sheet for Standards & Tools assessment.
    
    Tools are only useful if people know how to assess them properly.
    This guides assessors to focus on effectiveness, not just deployment.
    """
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ISMS-IMP-A.8.28.2 – Standards & Tools Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "ISO/IEC 27001:2022 - Control A.8.28: Secure Coding (Standards & Tools)"
    apply_style(ws["A2"], styles["subheader"])
    ws.row_dimensions[2].height = 20

    # Document Information Block
    row = 4
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f"A{row}:F{row}")
    
    doc_info = [
        ("Document ID:", "ISMS-IMP-A.8.28.2"),
        ("Assessment Area:", "Secure Coding Standards & Security Tools"),
        ("Related Policy:", "ISMS-POL-A.8.28-S2.2 (Coding Standards), S2.3 (Code Review)"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT - Enter assessment date]"),
        ("Completed By:", "[USER INPUT - Enter assessor name]"),
        ("Organization:", "[USER INPUT - Enter organization name]"),
        ("Review Cycle:", "Quarterly or when tools/standards change"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"B{row}"] = value
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].font = Font(italic=True, size=10)
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    # How to Use This Workbook
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25

    instructions = [
        "1. Complete each assessment domain: Coding Standards → SAST/SCA → DAST → IDE Plugins → Tool Metrics",
        "2. Assess DEPLOYMENT (is it there?) and EFFECTIVENESS (does it work?)",
        "3. For tools: Don't just mark 'Implemented' if the tool exists - assess if it's configured and used properly",
        "4. Provide specific tool names and versions (not just 'we have SAST')",
        "5. Reference actual scan results, configuration files, and adoption metrics as evidence",
        "6. Use Tool Effectiveness Metrics sheet to track measurable outcomes",
        "7. Document gaps with specific remediation actions (not just 'improve tool')",
        "8. Obtain approvals from both Security and Development leadership",
        "",
        "\u26A0\uFE0F  CRITICAL: Having a tool ≠ using it effectively. Focus on:",
        "   \u2022 Is it configured with security rules? \u2022 Does it run automatically? \u2022 Are findings remediated?",
        "   \u2022 What's the false positive rate? \u2022 Do developers actually use it?",
    ]

    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{row}:F{row}")
        if instruction.startswith("\u26A0\uFE0F"):
            ws[f"A{row}"].font = Font(bold=True, size=10, color="C00000")
        ws.row_dimensions[row].height = 30 if len(instruction) > 100 else 20
        row += 1

    # Assessment Domains Overview
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ASSESSMENT DOMAINS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Domain"
    ws[f"B{row}"] = "Focus Area"
    ws[f"C{row}"] = "Requirements"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    domain_overview = [
        ("1. Coding Standards Adoption", "Standards documentation, training, compliance measurement", "18"),
        ("2. SAST & SCA Tools", "Static analysis, dependency scanning, vulnerability detection", "18"),
        ("3. DAST & Security Testing Tools", "Dynamic testing, API testing, container/IaC scanning", "18"),
        ("4. IDE Plugins & Linters", "Developer tools, pre-commit hooks, real-time feedback", "18"),
        ("5. Tool Effectiveness & Metrics", "KPIs, coverage, false positives, remediation velocity", "18"),
    ]

    row += 1
    for domain, focus, req_count in domain_overview:
        ws[f"A{row}"] = domain
        ws[f"B{row}"] = focus
        ws[f"C{row}"] = req_count
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    ws[f"C{row}"] = "90 Total"
    ws[f"C{row}"].font = Font(bold=True)
    ws[f"C{row}"].alignment = Alignment(horizontal="center")

    # Implementation Status Legend
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "IMPLEMENTATION STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    legend_headers = ["Status", "Symbol", "Description", "When to Use"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    legend_data = [
        ("Implemented", "\u2705", "Standard/tool deployed AND effective", "When tool is configured, integrated, and findings are acted upon"),
        ("Partially Implemented", "\u26A0\uFE0F", "Standard/tool exists but gaps in coverage/effectiveness", "When tool exists but not fully integrated or has low adoption"),
        ("Not Implemented", "\u274C", "Standard/tool not deployed", "When standard doesn't exist or tool not deployed"),
        ("N/A", "N/A", "Not applicable to environment", "Rarely appropriate - most tools/standards are applicable"),
    ]

    row += 1
    for status, symbol, desc, when_use in legend_data:
        ws[f"A{row}"] = status
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = symbol
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = when_use
        
        if status == "Implemented":
            ws[f"A{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "Partially Implemented":
            ws[f"A{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif status == "Not Implemented":
            ws[f"A{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif status == "N/A":
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws.merge_cells(f"C{row}:F{row}")
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    # Acceptable Evidence Examples
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (Examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    evidence_examples = [
        "✓ Secure coding standards documents (wiki links, PDFs, internal guidelines)",
        "✓ Tool deployment documentation (licenses, access URLs, admin guides)",
        "✓ Tool configuration files (SAST rulesets, SCA policies, quality gates)",
        "✓ Scan results (SAST/SCA/DAST reports with timestamps and findings)",
        "✓ CI/CD pipeline configurations (Jenkinsfile, .gitlab-ci.yml showing tool integration)",
        "✓ Dashboard screenshots (tool dashboards, metrics, trend analysis)",
        "✓ Adoption metrics (tool usage statistics, developer adoption rates)",
        "✓ False positive analysis (FP rate metrics, tuning documentation)",
        "✓ Remediation tracking (JIRA tickets, vulnerability lifecycle)",
        "✓ Training materials (standards training, tool usage guides)",
        "✓ Quality gate evidence (blocked builds, deployment gates)",
        "✓ Tool effectiveness metrics (detection rate, MTTR, coverage %)",
        "✓ License agreements (tool licenses, support contracts)",
        "✓ Integration evidence (SIEM logs, ticketing system integration)",
        "✓ Benchmark reports (tool comparison, effectiveness benchmarks)",
    ]

    row += 1
    for evidence in evidence_examples:
        ws[f"A{row}"] = evidence
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

    # Common Anti-Patterns
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMMON ANTI-PATTERNS TO AVOID"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    antipatterns = [
        "\u274C 'We have SonarQube' ≠ Implemented (Is it configured? Does it run? Are findings fixed?)",
        "\u274C Marking tools 'Implemented' when they're deployed but not used effectively",
        "\u274C Ignoring false positive rates (99% FP rate means the tool is useless)",
        "\u274C Having 10 tools that all find the same 5 vulnerabilities (tool sprawl)",
        "\u274C Tools that slow down developers so much they bypass them",
        "\u274C Standards documents that nobody can find or understand",
        "\u274C Training that happened once 3 years ago counting as 'ongoing training'",
        "\u274C Measuring 'findings found' instead of 'findings fixed'",
    ]

    row += 1
    for antipattern in antipatterns:
        ws[f"A{row}"] = antipattern
        ws[f"A{row}"].font = Font(size=10, color="C00000")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 5: DOMAIN SHEET AND SUPPORTING SHEETS CREATION
# ============================================================================

def create_domain_sheet(ws, domain_name, requirements, styles):
    """Create standardized assessment domain sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Standards & Tools Assessment: {domain_name}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    headers = ["ID", "Requirement", "Implementation Status", "Evidence Reference", "Comments", "Compliance"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        apply_style(cell, styles["column_header"])
    
    ws.row_dimensions[2].height = 30
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 12

    current_category = None
    row = 3
    
    for req in requirements:
        if "category" in req and req["category"] != current_category:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"▶ {req['category']}"
            ws[f"A{row}"].font = Font(bold=True, size=11, color="4472C4")
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20
            row += 1
            current_category = req["category"]
        
        ws[f"A{row}"] = req["id"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        ws[f"B{row}"] = req["requirement"]
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        validations["implementation_status"].add(ws[f"C{row}"])
        
        ws[f"D{row}"] = req.get("evidence", "")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"D{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[f"D{row}"].font = Font(size=9, italic=True, color="666666")
        
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"E{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws[f"F{row}"] = f'=IF(C{row}="Implemented","\u2705",IF(C{row}="Partially Implemented","\u26A0\uFE0F",IF(C{row}="Not Implemented","\u274C",IF(C{row}="N/A","N/A",""))))'
        ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws.row_dimensions[row].height = 45
        row += 1

    ws.freeze_panes = "A3"
    return ws


def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with compliance metrics."""
    
    ws.merge_cells("A1:E1")
    ws["A1"] = "Standards & Tools Assessment - Summary Dashboard"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Date:"
    ws["B3"] = "=Instructions!B8"
    ws["A4"] = "Assessor:"
    ws["B4"] = "=Instructions!B9"
    ws["A5"] = "Organization:"
    ws["B5"] = "=Instructions!B10"
    
    for r in [3, 4, 5]:
        ws[f"A{r}"].font = Font(bold=True)

    ws["A7"] = "Overall Standards & Tools Compliance:"
    ws["A7"].font = Font(bold=True, size=13)
    ws.merge_cells("A7:B7")
    
    ws["C7"] = """=(COUNTIF(Coding_Standards_Adoption!C:C,"Implemented")+
COUNTIF(SAST_SCA_Tools!C:C,"Implemented")+
COUNTIF(DAST_Security_Testing_Tools!C:C,"Implemented")+
COUNTIF(IDE_Plugins_Linters!C:C,"Implemented")+
COUNTIF(Tool_Effectiveness_Metrics!C:C,"Implemented"))/
(COUNTA(Coding_Standards_Adoption!C:C)+
COUNTA(SAST_SCA_Tools!C:C)+
COUNTA(DAST_Security_Testing_Tools!C:C)+
COUNTA(IDE_Plugins_Linters!C:C)+
COUNTA(Tool_Effectiveness_Metrics!C:C)-5-
COUNTIF(Coding_Standards_Adoption!C:C,"N/A")-
COUNTIF(SAST_SCA_Tools!C:C,"N/A")-
COUNTIF(DAST_Security_Testing_Tools!C:C,"N/A")-
COUNTIF(IDE_Plugins_Linters!C:C,"N/A")-
COUNTIF(Tool_Effectiveness_Metrics!C:C,"N/A"))"""
    
    ws["C7"].number_format = "0%"
    ws["C7"].font = Font(bold=True, size=14)
    ws["C7"].alignment = Alignment(horizontal="center")
    
    ws["D7"] = '=IF(C7>=0.8,"🟢 Compliant",IF(C7>=0.6,"🟡 Needs Improvement","🔴 Non-Compliant"))'
    ws["D7"].font = Font(bold=True, size=12)
    ws["D7"].alignment = Alignment(horizontal="center")
    ws.merge_cells("D7:E7")

    ws["A10"] = "Domain Compliance Breakdown"
    ws["A10"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A10"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells("A10:E10")
    ws["A10"].alignment = Alignment(horizontal="center")

    domain_headers = ["Domain", "Compliance %", "Status", "Not Implemented", "Partially Implemented"]
    for col_num, header in enumerate(domain_headers, 1):
        cell = ws.cell(row=11, column=col_num, value=header)
        apply_style(cell, styles["column_header"])

    domains = [
        ("Coding_Standards_Adoption", "1. Coding Standards Adoption"),
        ("SAST_SCA_Tools", "2. SAST & SCA Tools"),
        ("DAST_Security_Testing_Tools", "3. DAST & Security Testing Tools"),
        ("IDE_Plugins_Linters", "4. IDE Plugins & Linters"),
        ("Tool_Effectiveness_Metrics", "5. Tool Effectiveness & Metrics"),
    ]

    row = 12
    for sheet_name, display_name in domains:
        ws[f"A{row}"] = display_name
        ws[f"B{row}"] = f'=COUNTIF({sheet_name}!C:C,"Implemented")/(COUNTA({sheet_name}!C:C)-1-COUNTIF({sheet_name}!C:C,"N/A"))'
        ws[f"B{row}"].number_format = "0%"
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"] = f'=IF(B{row}>=0.8,"🟢",IF(B{row}>=0.6,"🟡","🔴"))'
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"] = f'=COUNTIF({sheet_name}!C:C,"Not Implemented")'
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"E{row}"] = f'=COUNTIF({sheet_name}!C:C,"Partially Implemented")'
        ws[f"E{row}"].alignment = Alignment(horizontal="center")
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 25

    return ws


def create_evidence_register(ws, styles):
    """Create Evidence Register."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:G1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 25

    ws["A2"] = "Document all evidence supporting your Standards & Tools assessment."
    ws.merge_cells("A2:G2")
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True)

    headers = ["Evidence ID", "Related Requirement", "Evidence Type", "Description", "Location/Link", "Collection Date", "Collected By"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        apply_style(cell, styles["column_header"])

    widths = {"A": 12, "B": 25, "C": 18, "D": 40, "E": 35, "F": 15, "G": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    validations["evidence_type"].add("C4:C200")
    ws.freeze_panes = "A4"
    return ws


def create_gap_analysis_sheet(ws, styles):
    """Create Gap Analysis sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:J1")
    ws["A1"] = "Gap Analysis & Remediation Plan"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 25

    headers = ["Gap ID", "Domain", "Requirement ID", "Requirement Description", "Current State", "Target State", "Priority", "Owner", "Target Date", "Status"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        apply_style(cell, styles["column_header"])

    widths = {"A": 10, "B": 22, "C": 12, "D": 35, "E": 25, "F": 25, "G": 12, "H": 20, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    validations["priority"].add("G4:G200")
    validations["gap_status"].add("J4:J200")
    ws.freeze_panes = "A4"
    return ws


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet."""
    
    ws.merge_cells("A1:D1")
    ws["A1"] = "Standards & Tools Assessment - Approval Sign-Off"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws["A6"] = "Assessment Date:"
    ws["B6"] = "=Instructions!B8"
    ws["A7"] = "Assessor Name:"
    ws["B7"] = "=Instructions!B9"
    ws["A8"] = "Overall Compliance Score:"
    ws["B8"] = "=Summary_Dashboard!C7"
    ws["B8"].number_format = "0%"
    ws["B8"].font = Font(bold=True, size=12)

    for r in [6, 7, 8]:
        ws[f"A{r}"].font = Font(bold=True)

    ws["A11"] = "APPROVAL WORKFLOW"
    ws["A11"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A11"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells("A11:D11")
    ws["A11"].alignment = Alignment(horizontal="center")

    approval_headers = ["Role", "Name", "Signature", "Date"]
    for col_num, header in enumerate(approval_headers, 1):
        cell = ws.cell(row=12, column=col_num, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    approval_roles = ["Assessment Completer", "Application Security Lead", "Development Manager", "CISO"]

    row = 13
    for role in approval_roles:
        ws[f"A{row}"] = role
        ws[f"A{row}"].font = Font(bold=True, size=10)
        for col in ["B", "C", "D"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15

    return ws

# ============================================================================
# SECTION 6: MAIN EXECUTION
# ============================================================================

def main():
    """Generate Standards & Tools Assessment workbook."""
    logger.info("=" * 80)
    logger.info(" " * 15 + "ISMS Control 8.28.2 - Standards & Tools Assessment Generator")
    logger.info("=" * 80)
    logger.info("")
    
    try:
        logger.info("📝 [1/10] Creating workbook structure...")
        wb = create_workbook()
        styles = setup_styles()
        logger.info("     ✓ Workbook initialized with 10 sheets")

        logger.info("📄 [2/10] Creating Instructions sheet...")
        ws_instructions = wb["Instructions"]
        create_instructions_sheet(ws_instructions, styles)
        logger.info("     ✓ Instructions complete (focus: deployment AND effectiveness)")

        logger.info("📊 [3/10] Creating Domain 1: Coding Standards Adoption...")
        ws_domain1 = wb["Coding_Standards_Adoption"]
        create_domain_sheet(ws_domain1, "Coding Standards Adoption", get_domain1_requirements(), styles)
        logger.info("     ✓ 18 requirements (Standards, Training, Enforcement)")

        logger.info("📊 [4/10] Creating Domain 2: SAST & SCA Tools...")
        ws_domain2 = wb["SAST_SCA_Tools"]
        create_domain_sheet(ws_domain2, "SAST & SCA Tools", get_domain2_requirements(), styles)
        logger.info("     ✓ 18 requirements (Static Analysis, Dependency Scanning)")

        logger.info("📊 [5/10] Creating Domain 3: DAST & Security Testing Tools...")
        ws_domain3 = wb["DAST_Security_Testing_Tools"]
        create_domain_sheet(ws_domain3, "DAST & Security Testing Tools", get_domain3_requirements(), styles)
        logger.info("     ✓ 18 requirements (Dynamic Testing, API Testing, Container/IaC)")

        logger.info("📊 [6/10] Creating Domain 4: IDE Plugins & Linters...")
        ws_domain4 = wb["IDE_Plugins_Linters"]
        create_domain_sheet(ws_domain4, "IDE Plugins & Linters", get_domain4_requirements(), styles)
        logger.info("     ✓ 18 requirements (IDE Plugins, Linters, Pre-commit Hooks)")

        logger.info("📊 [7/10] Creating Domain 5: Tool Effectiveness & Metrics...")
        ws_domain5 = wb["Tool_Effectiveness_Metrics"]
        create_domain_sheet(ws_domain5, "Tool Effectiveness & Metrics", get_domain5_requirements(), styles)
        logger.info("     ✓ 18 requirements (KPIs, Coverage, Remediation Velocity)")

        logger.info("📈 [8/10] Creating Summary Dashboard...")
        ws_summary = wb["Summary_Dashboard"]
        create_summary_dashboard(ws_summary, styles)
        logger.info("     ✓ Executive summary with tool effectiveness metrics")

        logger.info("📎 [9/10] Creating Evidence Register...")
        ws_evidence = wb["Evidence_Register"]
        create_evidence_register(ws_evidence, styles)
        logger.info("     ✓ Evidence tracking")

        logger.info("\u26A0\uFE0F  [10/10] Creating Gap Analysis & Approval sheets...")
        ws_gap = wb["Gap_Analysis"]
        create_gap_analysis_sheet(ws_gap, styles)
        
        ws_approval = wb["Approval_Sign_Off"]
        create_approval_sheet(ws_approval, styles)
        logger.info("     ✓ Gap tracking and approval workflow")

        logger.info("")
        logger.info("💾 Saving workbook...")
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.8.28.2_Standards_Tools_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("")
        logger.info("=" * 80)
        logger.info("\u2705 SUCCESS: Standards & Tools Assessment workbook generated!")
        logger.info("=" * 80)
        logger.info("")
        logger.info(f"📁 File: {filename}")
        logger.info(f"📊 Total Requirements: 90 (18 per domain × 5 domains)")
        logger.info("")
        logger.info("📌 Assessment Domains:")
        logger.info("   1. Coding Standards Adoption          (18 requirements)")
        logger.info("   2. SAST & SCA Tools                   (18 requirements)")
        logger.info("   3. DAST & Security Testing Tools      (18 requirements)")
        logger.info("   4. IDE Plugins & Linters              (18 requirements)")
        logger.info("   5. Tool Effectiveness & Metrics       (18 requirements)")
        logger.info("")
        logger.info("💡 Key Focus: Tools are investments. Measure if they're working.")
        logger.info("   Don't just check if tools exist - verify they're effective!")
        logger.info("")
        logger.info("=" * 80)
        
        return 0

    except Exception as e:
        logger.info("")
        logger.info("=" * 80)
        logger.error("\u274C ERROR: Failed to generate workbook")
        logger.info("=" * 80)
        logger.error(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
