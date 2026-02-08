#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.28.5 - Compliance Dashboard (Master Consolidation)
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Dashboard: Executive Compliance Summary - All Assessment Domains

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment workbook structures, compliance
thresholds, and executive reporting requirements.

Key customization areas:
1. Input workbook schemas (MUST match actual assessment file structures)
2. Compliance scoring algorithms (adapt to your risk-based criteria)
3. Executive summary formatting (specific to stakeholder preferences)
4. Threshold definitions for Red/Amber/Green status (based on risk appetite)
5. Integration points with GRC platforms (if applicable to your environment)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
CRITICAL - READ BEFORE USE
--------------------------------------------------------------------------------

**This script consolidates data from FOUR distinct assessment workbooks.**

Each assessment workbook has DIFFERENT sheet structures, column layouts, and
data schemas. You MUST analyze the actual structure of each input workbook
before running this script.

**STEP 1**: Define actual workbook schemas (Lines ~200-400 - CUSTOMIZE section)
**STEP 2**: Validate workbook existence and schema compliance
**STEP 3**: Extract data using domain-specific logic (NOT copy-paste!)
**STEP 4**: Consolidate with proper error handling and data validation

**DO NOT assume all workbooks have identical structures.**
**DO NOT copy consolidation logic from other controls without adaptation.**

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an executive compliance dashboard that consolidates
assessment data from all four secure coding assessment domains into a single
integrated view for CISO/management oversight and audit readiness.

**Purpose:**
Provides executive-level visibility into secure coding compliance across the
entire software development lifecycle, enabling risk-based prioritization of
remediation efforts and demonstrating ISO 27001:2022 Control A.8.28 compliance.

**Input Workbooks (MUST exist before running this script):**
1. ISMS_A_8_28_1_SDLC_Assessment_YYYYMMDD.xlsx
   - SDLC security integration across all development phases
   - Security gate effectiveness and training compliance

2. ISMS_A_8_28_2_Standards_Tools_Assessment_YYYYMMDD.xlsx
   - Coding standards implementation per language/framework
   - SAST/SCA tool coverage and configuration compliance

3. ISMS_A_8_28_3_Code_Review_Testing_Assessment_YYYYMMDD.xlsx
   - Code review coverage and effectiveness metrics
   - Security testing coverage and vulnerability remediation rates

4. ISMS_A_8_28_4_Third_Party_OSS_Assessment_YYYYMMDD.xlsx
   - Third-party and OSS component risk management
   - Dependency vulnerability status and SBOM accuracy

**Generated Dashboard Structure:**
1. Executive Summary - High-level compliance status and key metrics
2. Domain Compliance - Individual assessment domain scores and trends
3. Critical Gaps - Prioritized list of high-risk gaps requiring remediation
4. Remediation Tracker - Action items with ownership and timelines
5. Trend Analysis - Compliance progress over time (multi-period comparison)
6. Evidence Summary - Consolidated audit evidence register
7. Risk Heat Map - Visual risk assessment across all domains
8. Recommendations - Prioritized improvement actions for management

**Key Features:**
- Automated data extraction from all four assessment workbooks
- Weighted compliance scoring based on risk severity
- Red/Amber/Green (RAG) status visualization
- Trend analysis comparing multiple assessment periods
- Gap prioritization based on risk and effort to remediate
- Executive-friendly charts and visualizations
- Audit-ready evidence consolidation
- Multi-stakeholder action tracking

**Consolidation Logic:**
- Overall Compliance Score: Weighted average of four domain scores
- Domain weighting: Configurable based on organizational risk profile
- Gap criticality: Based on vulnerability severity, exploitability, scope
- Trend analysis: Requires historical dashboard data for comparison

**Integration:**
This dashboard is the single source of truth for A.8.28 compliance status.
Used for:
- Executive/Board reporting
- Audit evidence presentation
- Risk management committee updates
- Continuous improvement tracking

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing and generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)
    - os (standard library - file operations)
    - json (standard library - optional, for config files)

Input Requirements:
    All FOUR assessment workbooks must exist and be completed:
    - ISMS_A_8_28_1_SDLC_Assessment_YYYYMMDD.xlsx
    - ISMS_A_8_28_2_Standards_Tools_Assessment_YYYYMMDD.xlsx
    - ISMS_A_8_28_3_Code_Review_Testing_Assessment_YYYYMMDD.xlsx
    - ISMS_A_8_28_4_Third_Party_OSS_Assessment_YYYYMMDD.xlsx
    
    All workbooks should use the SAME date suffix (YYYYMMDD) to ensure
    alignment and version consistency.

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Auto-detect latest assessment files in current directory
    python3 generate_a828_5_compliance_dashboard.py

Advanced Usage:
    # Specify input directory for assessment files
    python3 generate_a828_5_compliance_dashboard.py --input-dir /path/to/assessments
    
    # Specify date suffix to process specific assessment version
    python3 generate_a828_5_compliance_dashboard.py --date 20250124
    
    # Custom output location for dashboard
    python3 generate_a828_5_compliance_dashboard.py --output /path/to/dashboards
    
    # Include historical trend analysis (requires previous dashboards)
    python3 generate_a828_5_compliance_dashboard.py --trend-analysis
    
    # Specify custom domain weighting (JSON config file)
    python3 generate_a828_5_compliance_dashboard.py --weights config/weights.json
    
    # Validate input files without generating dashboard (dry-run)
    python3 generate_a828_5_compliance_dashboard.py --validate-only

Command-Line Options:
    --input-dir PATH       Directory containing assessment workbooks
    --output PATH          Output directory for dashboard (default: current dir)
    --date YYYYMMDD        Specific assessment date to process
    --trend-analysis       Include trend comparison with historical data
    --weights FILE         JSON file with custom domain weighting
    --validate-only        Validate input files without generating output
    --verbose              Enable detailed logging for troubleshooting

Output:
    File: ISMS_A_8_28_5_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Workflow Example:
    # Step 1: Validate all assessment files are present and complete
    python3 generate_a828_5_compliance_dashboard.py --validate-only --verbose
    
    # Step 2: Generate dashboard with trend analysis
    python3 generate_a828_5_compliance_dashboard.py --trend-analysis
    
    # Step 3: Review dashboard and distribute to stakeholders

Post-Generation Steps:
    1. Review executive summary for accuracy and completeness
    2. Validate compliance scores against raw assessment data
    3. Review critical gaps and confirm prioritization logic
    4. Assign remediation actions to responsible parties
    5. Schedule follow-up reviews based on remediation timelines
    6. Distribute dashboard to CISO, development leadership, audit team
    7. Archive dashboard for historical trend tracking
    8. Use dashboard in management review meetings and audit presentations

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    5 of 5 (Compliance Dashboard - Consolidation)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard Implementation Guide
    - ISMS-IMP-A.8.28.1: SDLC Integration Assessment (Input Domain 1)
    - ISMS-IMP-A.8.28.2: Standards & Tools Assessment (Input Domain 2)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Assessment (Input Domain 3)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Assessment (Input Domain 4)

Related Scripts:
    - generate_a828_1_sdlc_assessment.py (generates input workbook 1)
    - generate_a828_2_standards_tools.py (generates input workbook 2)
    - generate_a828_3_code_review_testing.py (generates input workbook 3)
    - generate_a828_4_third_party_oss.py (generates input workbook 4)
    - normalize_assessment_files_a828.py (pre-consolidation QA)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full dashboard framework per ISMS-IMP-A.8.28.5 specification
    - Consolidates all four assessment domains
    - Supports trend analysis and executive reporting
    - Integrated validation and error handling

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Schema Validation is Critical:**
This script MUST validate the structure of each input workbook before
attempting data extraction. Mismatched schemas will cause consolidation
failures or, worse, silent data corruption. Always validate first.

**Workbook Structure Differences:**
Each assessment domain has DIFFERENT sheet names, column layouts, and data
types. The consolidation logic MUST be domain-specific. DO NOT assume uniform
structure across all input workbooks.

**Error Handling Philosophy:**
Fail fast with clear error messages. If an input workbook is missing, corrupted,
or has unexpected structure, STOP and report the issue. Don't attempt to
"work around" missing data - that creates false compliance reports.

**Compliance Scoring Methodology:**
Default scoring uses weighted average of domain scores. Weighting reflects
risk impact of each domain. Customize weights based on your risk assessment:
- SDLC Integration: Foundation for all other controls
- Standards & Tools: Preventive controls effectiveness
- Code Review & Testing: Detective controls effectiveness
- Third-Party & OSS: Supply chain risk exposure

**Audit Considerations:**
This dashboard is PRIMARY audit evidence for A.8.28 compliance. Ensure:
- All calculations are transparent and traceable
- Source data references are accurate
- Evidence links point to actual audit artifacts
- Dashboard is signed off by appropriate stakeholders

Auditors will validate dashboard data against source assessments.

**Data Protection:**
Dashboard contains aggregated security posture information:
- Overall vulnerability exposure across applications
- Critical gaps and security weaknesses
- Remediation timelines and resource allocation
- Historical compliance trends

Handle as CONFIDENTIAL per your organization's data classification.

**Maintenance and Updates:**
Regenerate dashboard:
- Quarterly: Regular compliance status updates
- Monthly: If high-priority remediation in progress
- Ad-hoc: When requested for audit or executive review
- After major changes: New applications, framework updates, tool changes

Keep historical dashboards for trend analysis (minimum 12 months).

**Quality Assurance:**
Before distributing dashboard to executives or auditors:
- Validate all calculations manually (spot-check against source data)
- Review critical gaps list for accuracy and completeness
- Confirm remediation timelines are realistic
- Verify chart data matches summary tables

Have a second AppSec team member review before publication.

**Regulatory Alignment:**
Dashboard should reflect regulatory compliance requirements where applicable:
- PCI DSS v4.0.1: Secure coding compliance status (Requirement 6)
- HIPAA: Application security posture
- SOC 2: Secure development controls effectiveness
- ISO 27001: Evidence for Control A.8.28 implementation

Customize executive summary to highlight regulatory-specific compliance.

**Don't Fool Yourself - Feynman Principle:**
A green dashboard doesn't mean you're secure. Key reality checks:
- Are assessment scores based on actual evidence, or self-reported claims?
- Do "completed" remediation items actually fix the security issue?
- Are you measuring what matters, or what's easy to measure?
- If you were attacking your own applications, would this dashboard scare you?

The dashboard is only as good as the underlying assessments. Garbage in,
garbage out. Focus on honest, evidence-based assessment completion.

**Executive Communication:**
Dashboard should answer executive questions:
- "Are we secure?" → Overall compliance score with context
- "Where are the biggest risks?" → Critical gaps section
- "Are we getting better or worse?" → Trend analysis
- "What do we need to fix first?" → Prioritized remediation tracker
- "When will we be compliant?" → Remediation timeline projections

Make it executive-friendly: more charts, less text; highlight actions, not
just status; show trends, not just snapshots.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.28.5"
WORKBOOK_NAME = "Compliance Summary Dashboard"
CONTROL_ID = "A.8.28"
CONTROL_NAME = "Secure Coding"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches markdown specification
    sheets = [
        "Executive Dashboard",
        "Gap Analysis",
        "Risk Register",
        "Remediation Roadmap",
        "KPIs & Metrics",
        "Evidence Register",
        "Action Items & Follow-up",
        "Audit & Compliance Log",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "critical_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_yellow": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: EXECUTIVE DASHBOARD
# ============================================================================

def create_executive_dashboard(ws, styles):
    """Create Executive Dashboard with external workbook links and KPIs."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "ISMS-IMP-A.8.28.5 – Secure Coding Compliance Summary Dashboard\n"
        "ISO/IEC 27001:2022 - Control A.8.28: Secure Coding - Executive Overview"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # ---------- DOCUMENT INFORMATION ----------
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.28.5"),
        ("Report Type", "Compliance Summary Dashboard"),
        ("Related Policy", "ISMS-POL-A.8.28 (Secure Coding)"),
        ("Version", "1.0"),
        ("Report Date", ""),
        ("Reporting Period", ""),
        ("Prepared By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
        ("Last Updated", "=TODAY()"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        elif "=" in str(value):
            ws[f"B{row}"].font = Font(color="0000FF")
        row += 1

    # ---------- OVERALL COMPLIANCE STATUS ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "OVERALL CRYPTOGRAPHY COMPLIANCE STATUS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    # Compliance Scorecard
    row += 1
    scorecard_headers = ["Metric", "Score", "Target", "Status"]
    for col_idx, header in enumerate(scorecard_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    scorecard_start = row + 1
    scorecard_metrics = [
        ("Overall Compliance Rate", "=AVERAGE(C25:C28)", "≥95%", ""),
        ("Critical Gaps", "=COUNTIFS('Gap Analysis'!I8:I207,\"Critical\",'Gap Analysis'!R8:R207,\"<>Closed\")", "0", ""),
        ("High-Risk Items", "=COUNTIFS('Risk Register'!N20:N119,\">=15\",'Risk Register'!Q20:Q119,\"<>Closed\")", "≤5", ""),
        ("Remediation Progress", "=COUNTIF('Remediation Roadmap'!H37:H236,\"Completed\")/COUNTA('Remediation Roadmap'!H37:H236)*100&\"%\"", "≥80%", ""),
    ]

    row += 1
    for metric, formula, target, status in scorecard_metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        ws.cell(row=row, column=2, value=formula).font = Font(bold=True, color="0000FF")
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]  # Manual traffic light
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    # ---------- COMPLIANCE BY ASSESSMENT AREA ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "COMPLIANCE BY ASSESSMENT AREA"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    area_headers = [
        "Assessment Area",
        "Source Document",
        "Overall Compliance %",
        "Total Items",
        "Compliant",
        "Partial",
        "Non-Compliant",
        "Trend"
    ]
    
    for col_idx, header in enumerate(area_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Column widths
    widths = [25, 22, 18, 12, 12, 12, 14, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows with EXTERNAL WORKBOOK LINKS
    assessment_areas = [
        ("SDLC Assessment", "ISMS-IMP-A.8.28.1.xlsx"),
        ("Standards & Tools Assessment", "ISMS-IMP-A.8.28.2.xlsx"),
        ("Code Review & Testing Assessment", "ISMS-IMP-A.8.28.3.xlsx"),
        ("Third-Party & OSS Assessment", "ISMS-IMP-A.8.28.4.xlsx"),
    ]

    row += 1
    area_start_row = row
    for area_name, source_doc in assessment_areas:
        ws.cell(row=row, column=1, value=area_name)
        ws.cell(row=row, column=2, value=source_doc)
        
        # EXTERNAL LINKS - these formulas reference normalized source workbooks
        ws.cell(row=row, column=3, value=f"=[{source_doc}]'Summary Dashboard'!G8")  # Overall compliance %
        ws.cell(row=row, column=4, value=f"=[{source_doc}]'Summary Dashboard'!B8")  # Total items
        ws.cell(row=row, column=5, value=f"=[{source_doc}]'Summary Dashboard'!C8")  # Compliant
        ws.cell(row=row, column=6, value=f"=[{source_doc}]'Summary Dashboard'!D8")  # Partial
        ws.cell(row=row, column=7, value=f"=[{source_doc}]'Summary Dashboard'!E8")  # Non-Compliant
        
        # Trend - manual entry with dropdown
        ws.cell(row=row, column=8).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles["border"]
        
        row += 1

    # Trend dropdown validation
    dv_trend = DataValidation(type="list", formula1='"↑ Improved,→ No Change,↓ Declined"', allow_blank=False)
    ws.add_data_validation(dv_trend)
    for r in range(area_start_row, area_start_row + 4):
        dv_trend.add(ws.cell(row=r, column=8))

    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value="(Aggregated)").font = Font(italic=True)
    ws.cell(row=row, column=3, value=f"=AVERAGE(C{area_start_row}:C{row-1})").font = Font(bold=True, color="0000FF", size=12)
    ws.cell(row=row, column=4, value=f"=SUM(D{area_start_row}:D{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E{area_start_row}:E{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{area_start_row}:F{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"=SUM(G{area_start_row}:G{row-1})").font = Font(bold=True)
    
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = styles["border"]

    # ---------- CRITICAL METRICS SUMMARY ----------
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "KEY PERFORMANCE INDICATORS (KPIs)"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    kpi_headers = ["KPI", "Current Value", "Target", "Status", "Last Quarter", "Change"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    kpi_metrics = [
        ("Cryptographic Controls Implemented", "", "[Target]", "", "", ""),
        ("Systems with Encryption at Rest", "", "100%", "", "", ""),
        ("Systems with Encryption in Transit", "", "100%", "", "", ""),
        ("MFA Coverage (All Users)", "", "100%", "", "", ""),
        ("MFA Coverage (Admins)", "", "100%", "", "", ""),
        ("Certificate Expiry Compliance", "", "100%", "", "", ""),
        ("Key Rotation Compliance", "", "≥95%", "", "", ""),
        ("Password Hashing Compliance", "", "100%", "", "", ""),
        ("SSO Coverage", "", "≥80%", "", "", ""),
        ("Open Critical Gaps", f"=COUNTIFS('Gap Analysis'!I8:I207,\"Critical\",'Gap Analysis'!R8:R207,\"<>Closed\")", "0", "", "", ""),
    ]

    row += 1
    for kpi_name, current, target, status, last_q, change in kpi_metrics:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=2, value=current)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]  # Status - manual
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]  # Last Quarter - manual
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]  # Change - manual
        
        if "=" in str(current):
            ws.cell(row=row, column=2).font = Font(color="0000FF")
        else:
            ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = styles["border"]
        
        row += 1

    # ---------- TOP 5 CRITICAL ISSUES ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "TOP 5 CRITICAL SECURITY GAPS"
    ws[f"A{row}"].font = styles["critical_header"]["font"]
    ws[f"A{row}"].fill = styles["critical_header"]["fill"]
    ws[f"A{row}"].alignment = styles["critical_header"]["alignment"]

    row += 1
    critical_headers = ["Rank", "Issue Description", "Assessment Area", "Risk Level", "Systems Affected", "Target Date", "Owner", "Status"]
    for col_idx, header in enumerate(critical_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for rank in range(1, 6):
        ws.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal="center")
        for col in range(2, 9):
            ws.cell(row=row, column=col).fill = styles["input_cell"]["fill"]
            ws.cell(row=row, column=col).border = styles["border"]
        row += 1

    # ---------- EXECUTIVE SUMMARY ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "EXECUTIVE SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    summary_fields = [
        ("Assessment Period:", ""),
        ("Overall Compliance Status:", "=IF(C21>=95%,\"Excellent\",IF(C21>=90%,\"Good\",IF(C21>=80%,\"Adequate\",IF(C21>=70%,\"Needs Improvement\",\"Critical\"))))"),
        ("Security Posture:", "=IF(C21>=90%,\"Strong\",IF(C21>=75%,\"Adequate\",\"Weak\"))"),
    ]

    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = value
        if "=" in str(value):
            ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        else:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "Key Achievements This Period:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws.merge_cells(f"A{row}:H{row+4}")
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{row}"].border = styles["border"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row += 6
    ws[f"A{row}"] = "Major Concerns:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws.merge_cells(f"A{row}:H{row+4}")
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{row}"].border = styles["border"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row += 6
    ws[f"A{row}"] = "Recommended Executive Actions:"
    ws[f"A{row}"].font = Font(bold=True)
    
    for i in range(1, 4):
        row += 1
        ws[f"A{row}"] = f"{i}."
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:H{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    row += 2
    ws[f"A{row}"] = "Budget Impact:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    budget_fields = [
        ("Total Estimated Remediation Budget:", "=SUM('Remediation Roadmap'!R37:R236)"),
        ("Approved Budget:", ""),
        ("Budget Gap:", "=B[row]-B[row-1]"),
    ]
    
    budget_start = row
    for label, formula in budget_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        resolved_formula = formula.replace("[row-1]", str(row-1)).replace("[row]", str(row))
        ws[f"B{row}"] = resolved_formula
        if "=" in formula:
            ws[f"B{row}"].font = Font(color="0000FF")
            ws[f"B{row}"].number_format = '"$"#,##0.00'
        else:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].number_format = '"$"#,##0.00'
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws.freeze_panes = "A4"


# ============================================================================
# END OF PART 2A
# ============================================================================

# ============================================================================
# SECTION 3: GAP ANALYSIS
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create comprehensive gap analysis sheet with summary statistics."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:T1")
    ws["A1"] = (
        "COMPREHENSIVE GAP ANALYSIS - CRYPTOGRAPHIC CONTROLS\n"
        "Consolidated view of all compliance gaps across assessment areas"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- SUMMARY STATISTICS ----------
    ws["A3"] = "Gap Summary Statistics"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    summary_headers = [
        "Assessment Area",
        "Total Gaps",
        "Critical",
        "High",
        "Medium",
        "Low",
        "Compliance %",
        "Risk Score"
    ]
    
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = ["Data Transmission", "Data Storage", "Authentication", "Key Management"]
    
    row += 1
    summary_start = row
    for area in areas:
        ws.cell(row=row, column=1, value=area)
        # Formulas count gaps by area and severity from detailed register below
        ws.cell(row=row, column=2, value=f'=COUNTIF(B12:B211,"{area}")')
        ws.cell(row=row, column=3, value=f'=COUNTIFS(B12:B211,"{area}",I12:I211,"Critical")')
        ws.cell(row=row, column=4, value=f'=COUNTIFS(B12:B211,"{area}",I12:I211,"High")')
        ws.cell(row=row, column=5, value=f'=COUNTIFS(B12:B211,"{area}",I12:I211,"Medium")')
        ws.cell(row=row, column=6, value=f'=COUNTIFS(B12:B211,"{area}",I12:I211,"Low")')
        ws.cell(row=row, column=7).fill = styles["input_cell"]["fill"]  # Manual entry
        # Risk Score: (Critical*10 + High*7 + Medium*4 + Low*1) / Total
        ws.cell(row=row, column=8, value=f'=IF(B{row}=0,0,(C{row}*10+D{row}*7+E{row}*4+F{row}*1)/B{row})')
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles["border"]
            if col >= 2 and col != 7:
                ws.cell(row=row, column=col).font = Font(color="0000FF")
        
        row += 1

    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 8):
        ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{summary_start}:{get_column_letter(col)}{row-1})")
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = styles["border"]
    
    ws.cell(row=row, column=8, value=f"=AVERAGE(H{summary_start}:H{row-1})")
    ws.cell(row=row, column=8).font = Font(bold=True)
    ws.cell(row=row, column=8).border = styles["border"]

    # Column widths for summary
    widths_summary = [22, 12, 10, 10, 10, 10, 14, 12]
    for col_idx, width in enumerate(widths_summary, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---------- DETAILED GAP REGISTER ----------
    row += 3
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "DETAILED GAP REGISTER"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    detail_headers = [
        "Gap ID",
        "Assessment Area",
        "Source Document",
        "System/Application",
        "Gap Description",
        "Policy Reference",
        "Current State",
        "Required State",
        "Risk Level",
        "Impact",
        "Likelihood",
        "Risk Score",
        "Data Classification",
        "Systems Affected",
        "Users Impacted",
        "Business Impact",
        "Compensating Controls",
        "Status",
        "Exception ID",
        "Risk ID"
    ]
    
    widths_detail = [12, 22, 20, 28, 40, 20, 30, 30, 14, 12, 12, 10, 18, 12, 12, 30, 30, 14, 14, 14]
    
    for col_idx, (header, width) in enumerate(zip(detail_headers, widths_detail), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Create validations
    dv_area = DataValidation(
        type="list",
        formula1='"Data Transmission,Data Storage,Authentication,Key Management"',
        allow_blank=False
    )
    ws.add_data_validation(dv_area)

    dv_source = DataValidation(
        type="list",
        formula1='"ISMS-IMP-A.8.28.1,ISMS-IMP-A.8.28.2,ISMS-IMP-A.8.28.3,ISMS-IMP-A.8.28.4"',
        allow_blank=False
    )
    ws.add_data_validation(dv_source)

    dv_risk_level = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_risk_level)

    dv_impact = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_impact)

    dv_likelihood = DataValidation(
        type="list",
        formula1='"Very High,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_likelihood)

    dv_classification = DataValidation(
        type="list",
        formula1='"Restricted,Confidential,Internal,Public"',
        allow_blank=False
    )
    ws.add_data_validation(dv_classification)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Closed,Accepted Risk"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Data entry rows (200 rows for gaps)
    row += 1
    data_start = row
    for r in range(row, row + 200):
        # Gap ID auto-increment
        ws.cell(row=r, column=1, value=f"GAP-{r-row+1:03d}").font = Font(color="808080")
        
        # Input cells
        for col in range(2, 21):
            cell = ws.cell(row=r, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        
        # Apply validations
        dv_area.add(ws.cell(row=r, column=2))       # B: Assessment Area
        dv_source.add(ws.cell(row=r, column=3))     # C: Source Document
        dv_risk_level.add(ws.cell(row=r, column=9)) # I: Risk Level
        dv_impact.add(ws.cell(row=r, column=10))    # J: Impact
        dv_likelihood.add(ws.cell(row=r, column=11))# K: Likelihood
        dv_classification.add(ws.cell(row=r, column=13)) # M: Data Classification
        dv_status.add(ws.cell(row=r, column=18))    # R: Status
        
        # Risk Score formula (column L = 12)
        # Simple scoring: Critical=25, High=20, Medium=15, Low=10
        ws.cell(row=r, column=12, value=f'=IF(OR(ISBLANK(J{r}),ISBLANK(K{r})),"",'
                                               f'IF(AND(J{r}="Critical",K{r}="Very High"),25,'
                                               f'IF(AND(J{r}="Critical",K{r}="High"),20,'
                                               f'IF(AND(J{r}="High",K{r}="Very High"),20,'
                                               f'IF(AND(J{r}="High",K{r}="High"),16,'
                                               f'IF(AND(J{r}="Critical",K{r}="Medium"),15,'
                                               f'IF(AND(J{r}="Medium",K{r}="High"),12,'
                                               f'IF(AND(J{r}="High",K{r}="Medium"),12,10))))))))')
        ws.cell(row=r, column=12).font = Font(color="0000FF")

    ws.freeze_panes = "A12"


# ============================================================================
# SECTION 4: RISK REGISTER
# ============================================================================

def create_risk_register(ws, styles):
    """Create comprehensive risk register with heat map and treatment plans."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:U1")
    ws["A1"] = (
        "CRYPTOGRAPHY RISK REGISTER\n"
        "Consolidated risk assessment for cryptographic control failures"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- RISK SUMMARY DASHBOARD ----------
    ws["A3"] = "Risk Summary Dashboard"
    ws["A3"].font = Font(bold=True, size=11)

    # Risk Distribution
    row = 4
    ws[f"A{row}"] = "Risk Distribution"
    ws[f"A{row}"].font = Font(bold=True)
    
    row += 1
    dist_headers = ["Risk Category", "Count", "Percentage", "Risk Exposure"]
    for col_idx, header in enumerate(dist_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    categories = ["Critical", "High", "Medium", "Low"]
    row += 1
    dist_start = row
    
    for category in categories:
        ws.cell(row=row, column=1, value=category)
        # Count from detailed register (rows 20+)
        ws.cell(row=row, column=2, value=f'=COUNTIF(N20:N119,">={20 if category=="Critical" else 15 if category=="High" else 10 if category=="Medium" else 0}")')
        ws.cell(row=row, column=3, value=f'=IF(B{dist_start+3}=0,"0%",B{row}/B{dist_start+3}*100&"%")')
        ws.cell(row=row, column=4, value=f'=SUMIF(N20:N119,">={20 if category=="Critical" else 15 if category=="High" else 10 if category=="Medium" else 0}",N20:N119)')
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = styles["border"]
            if col > 1:
                ws.cell(row=row, column=col).font = Font(color="0000FF")
        row += 1

    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{dist_start}:B{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"=SUM(D{dist_start}:D{row-1})").font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = styles["border"]

    # Risk By Assessment Area
    row += 2
    ws[f"A{row}"] = "Risk By Assessment Area"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    area_headers = ["Assessment Area", "Critical", "High", "Medium", "Low", "Total Risk Exposure", "Avg Risk Score"]
    for col_idx, header in enumerate(area_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = ["Data Transmission", "Data Storage", "Authentication", "Key Management"]
    row += 1
    
    for area in areas:
        ws.cell(row=row, column=1, value=area)
        # Link to Gap Analysis sheet for area-based counts
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6, value=f"=B{row}*25+C{row}*20+D{row}*15+E{row}*10")
        ws.cell(row=row, column=6).font = Font(color="0000FF")
        ws.cell(row=row, column=7, value=f"=IF(SUM(B{row}:E{row})=0,0,F{row}/SUM(B{row}:E{row}))")
        ws.cell(row=row, column=7).font = Font(color="0000FF")
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = styles["border"]
        row += 1

    # Column widths
    widths_summary = [22, 12, 12, 12, 12, 20, 16]
    for col_idx, width in enumerate(widths_summary, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---------- DETAILED RISK REGISTER ----------
    row += 2
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "DETAILED RISK REGISTER"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    detail_headers = [
        "Risk ID",
        "Gap ID (Link)",
        "Risk Category",
        "Risk Description",
        "Threat Source",
        "Vulnerability",
        "Existing Controls",
        "Control Effectiveness",
        "Inherent Likelihood",
        "Inherent Impact",
        "Inherent Risk Score",
        "Residual Likelihood",
        "Residual Impact",
        "Residual Risk Score",
        "Risk Response",
        "Risk Owner",
        "Risk Status",
        "Target Risk Score",
        "Review Date",
        "Last Review Date",
        "Next Review Due"
    ]
    
    widths_detail = [12, 12, 22, 40, 25, 30, 30, 16, 14, 14, 12, 14, 14, 12, 18, 20, 14, 12, 14, 14, 14]
    
    for col_idx, (header, width) in enumerate(zip(detail_headers, widths_detail), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Create validations
    dv_category = DataValidation(
        type="list",
        formula1='"Cryptographic Weakness,Key Management,Authentication Failure,Data Exposure,Compliance Violation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_threat = DataValidation(
        type="list",
        formula1='"External Attacker,Insider Threat,System Failure,Human Error,Vendor/Third Party"',
        allow_blank=False
    )
    ws.add_data_validation(dv_threat)

    dv_effectiveness = DataValidation(
        type="list",
        formula1='"Effective,Partially Effective,Ineffective,None"',
        allow_blank=False
    )
    ws.add_data_validation(dv_effectiveness)

    dv_likelihood = DataValidation(
        type="list",
        formula1='"Very High,High,Medium,Low,Very Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_likelihood)

    dv_impact = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_impact)

    dv_response = DataValidation(
        type="list",
        formula1='"Mitigate,Accept,Transfer,Avoid"',
        allow_blank=False
    )
    ws.add_data_validation(dv_response)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,Monitoring,Closed,Accepted"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Data entry rows (100 rows for risks)
    row += 1
    data_start = row
    
    for r in range(row, row + 100):
        # Risk ID auto-increment
        ws.cell(row=r, column=1, value=f"RISK-{r-row+1:03d}").font = Font(color="808080")
        
        # Input cells
        for col in range(2, 22):
            cell = ws.cell(row=r, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        
        # Apply validations
        dv_category.add(ws.cell(row=r, column=3))     # C: Risk Category
        dv_threat.add(ws.cell(row=r, column=5))       # E: Threat Source
        dv_effectiveness.add(ws.cell(row=r, column=8))# H: Control Effectiveness
        dv_likelihood.add(ws.cell(row=r, column=9))   # I: Inherent Likelihood
        dv_impact.add(ws.cell(row=r, column=10))      # J: Inherent Impact
        dv_likelihood.add(ws.cell(row=r, column=12))  # L: Residual Likelihood
        dv_impact.add(ws.cell(row=r, column=13))      # M: Residual Impact
        dv_response.add(ws.cell(row=r, column=15))    # O: Risk Response
        dv_status.add(ws.cell(row=r, column=17))      # Q: Risk Status
        
        # Inherent Risk Score (column K = 11)
        ws.cell(row=r, column=11, value=f'=IF(OR(ISBLANK(I{r}),ISBLANK(J{r})),"",'
                                               f'IF(AND(J{r}="Critical",I{r}="Very High"),25,'
                                               f'IF(AND(J{r}="Critical",I{r}="High"),20,'
                                               f'IF(AND(J{r}="High",I{r}="Very High"),20,'
                                               f'IF(AND(J{r}="High",I{r}="High"),16,10)))))')
        ws.cell(row=r, column=11).font = Font(color="0000FF")
        
        # Residual Risk Score (column N = 14)
        ws.cell(row=r, column=14, value=f'=IF(OR(ISBLANK(L{r}),ISBLANK(M{r})),"",'
                                               f'IF(AND(M{r}="Critical",L{r}="Very High"),25,'
                                               f'IF(AND(M{r}="Critical",L{r}="High"),20,'
                                               f'IF(AND(M{r}="High",L{r}="Very High"),20,'
                                               f'IF(AND(M{r}="High",L{r}="High"),16,10)))))')
        ws.cell(row=r, column=14).font = Font(color="0000FF")
        
        # Next Review Due (column U = 21) - auto-calculate +90 days from Review Date
        ws.cell(row=r, column=21, value=f'=IF(ISBLANK(S{r}),"",S{r}+90)')
        ws.cell(row=r, column=21).font = Font(color="0000FF")

    ws.freeze_panes = "A20"


# ============================================================================
# END OF PART 2B
# ============================================================================

# ============================================================================
# SECTION 5: REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap(ws, styles):
    """Create remediation roadmap with progress tracking and timeline."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:Z1")
    ws["A1"] = (
        "CRYPTOGRAPHY REMEDIATION ROADMAP\n"
        "Action plan for closing identified gaps and reducing risk"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- ROADMAP SUMMARY ----------
    ws["A3"] = "Roadmap Summary - Overall Progress"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    summary_headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    summary_metrics = [
        ("Total Remediation Items", "=COUNTA(E37:E236)", "N/A", ""),
        ("Completed", '=COUNTIF(H37:H236,"Completed")', "-", ""),
        ("In Progress", '=COUNTIF(H37:H236,"In Progress")', "-", ""),
        ("Not Started", '=COUNTIF(H37:H236,"Not Started")', "-", ""),
        ("Overdue", '=COUNTIFS(M37:M236,"<"&TODAY(),H37:H236,"<>Completed")', "0", ""),
        ("On Schedule", '=COUNTIFS(M37:M236,">="&TODAY(),H37:H236,"<>Completed")', "100%", ""),
        ("Total Budget Required", "=SUM(R37:R236)", "-", ""),
        ("Budget Allocated", "", "-", ""),
        ("Budget Utilized", "=SUM(T37:T236)", "-", ""),
    ]

    row += 1
    for metric, formula, target, status in summary_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        
        if "=" in str(formula):
            ws.cell(row=row, column=2).font = Font(color="0000FF")
            if "SUM" in str(formula):
                ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        else:
            ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
            if metric == "Budget Allocated":
                ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = styles["border"]
        row += 1

    # Progress by Assessment Area
    row += 2
    ws[f"A{row}"] = "Progress by Assessment Area"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    area_headers = [
        "Assessment Area",
        "Total Items",
        "Completed",
        "In Progress",
        "Not Started",
        "% Complete",
        "On-Time %"
    ]
    
    for col_idx, header in enumerate(area_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = ["Data Transmission", "Data Storage", "Authentication", "Key Management"]
    row += 1
    area_start = row
    
    for area in areas:
        ws.cell(row=row, column=1, value=area)
        ws.cell(row=row, column=2, value=f'=COUNTIF(D37:D236,"{area}")')
        ws.cell(row=row, column=3, value=f'=COUNTIFS(D37:D236,"{area}",H37:H236,"Completed")')
        ws.cell(row=row, column=4, value=f'=COUNTIFS(D37:D236,"{area}",H37:H236,"In Progress")')
        ws.cell(row=row, column=5, value=f'=COUNTIFS(D37:D236,"{area}",H37:H236,"Not Started")')
        ws.cell(row=row, column=6, value=f'=IF(B{row}=0,"0%",C{row}/B{row}*100&"%")')
        ws.cell(row=row, column=7).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = styles["border"]
            if col >= 2 and col <= 6:
                ws.cell(row=row, column=col).font = Font(color="0000FF")
        row += 1

    # Column widths for summary
    widths_summary = [22, 14, 14, 14, 14, 14, 14]
    for col_idx, width in enumerate(widths_summary, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---------- GANTT CHART PLACEHOLDER ----------
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "REMEDIATION TIMELINE BY QUARTER"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    ws[f"A{row}"] = "(Gantt chart visualization - use conditional formatting or insert chart manually)"
    ws[f"A{row}"].font = Font(italic=True, color="808080")

    # ---------- DETAILED REMEDIATION REGISTER ----------
    row += 3
    ws.merge_cells(f"A{row}:Z{row}")
    ws[f"A{row}"] = "DETAILED REMEDIATION REGISTER"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    detail_headers = [
        "Remediation ID",
        "Gap ID (Link)",
        "Risk ID (Link)",
        "Assessment Area",
        "Remediation Action",
        "Action Type",
        "Priority",
        "Current Status",
        "% Complete",
        "Responsible Person",
        "Supporting Team",
        "Start Date",
        "Target Date",
        "Actual Completion",
        "Days Remaining",
        "On Schedule",
        "Dependencies",
        "Budget Required",
        "Budget Approved",
        "Actual Cost",
        "Vendor/Resource",
        "Success Criteria",
        "Verification Method",
        "Completion Evidence",
        "Notes/Comments",
        "Last Updated"
    ]
    
    widths_detail = [14, 12, 12, 22, 40, 20, 12, 16, 10, 20, 20, 14, 14, 14, 10, 12, 30, 14, 14, 14, 20, 30, 25, 28, 30, 14]
    
    for col_idx, (header, width) in enumerate(zip(detail_headers, widths_detail), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Create validations
    dv_area = DataValidation(
        type="list",
        formula1='"Data Transmission,Data Storage,Authentication,Key Management"',
        allow_blank=False
    )
    ws.add_data_validation(dv_area)

    dv_action_type = DataValidation(
        type="list",
        formula1='"Technical Implementation,Process Change,Policy Update,Training,Tool Procurement,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_action_type)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,Planning,In Progress,Testing,Completed,On Hold,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_budget_approved = DataValidation(
        type="list",
        formula1='"Yes,No,Pending"',
        allow_blank=False
    )
    ws.add_data_validation(dv_budget_approved)

    # Data entry rows (200 rows for remediation items)
    row += 1
    data_start = row
    
    for r in range(row, row + 200):
        # Remediation ID auto-increment
        ws.cell(row=r, column=1, value=f"REM-{r-row+1:03d}").font = Font(color="808080")
        
        # Input cells
        for col in range(2, 27):
            cell = ws.cell(row=r, column=col)
            
            # Skip formula columns
            if col not in [9, 15, 16, 26]:  # % Complete, Days Remaining, On Schedule, Last Updated
                cell.fill = styles["input_cell"]["fill"]
            
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        
        # Apply validations
        dv_area.add(ws.cell(row=r, column=4))         # D: Assessment Area
        dv_action_type.add(ws.cell(row=r, column=6))  # F: Action Type
        dv_priority.add(ws.cell(row=r, column=7))     # G: Priority
        dv_status.add(ws.cell(row=r, column=8))       # H: Current Status
        dv_budget_approved.add(ws.cell(row=r, column=19)) # S: Budget Approved
        
        # % Complete (column I = 9) - manual entry but validate 0-100
        ws.cell(row=r, column=9).fill = styles["input_cell"]["fill"]
        ws.cell(row=r, column=9).number_format = '0"%"'
        
        # Budget fields - currency format
        ws.cell(row=r, column=18).number_format = '"$"#,##0.00'  # R: Budget Required
        ws.cell(row=r, column=20).number_format = '"$"#,##0.00'  # T: Actual Cost
        
        # Days Remaining (column O = 15)
        ws.cell(row=r, column=15, value=f'=IF(H{r}="Completed",0,IF(ISBLANK(M{r}),"",M{r}-TODAY()))')
        ws.cell(row=r, column=15).font = Font(color="0000FF")
        ws.cell(row=r, column=15).fill = styles["formula_cell"]["fill"]
        
        # On Schedule (column P = 16)
        ws.cell(row=r, column=16, value=f'=IF(H{r}="Completed","Completed",'
                                              f'IF(ISBLANK(O{r}),"",IF(O{r}<0,"Overdue",'
                                              f'IF(O{r}<30,"At Risk","On Track"))))')
        ws.cell(row=r, column=16).font = Font(color="0000FF")
        ws.cell(row=r, column=16).fill = styles["formula_cell"]["fill"]
        
        # Last Updated (column Z = 26) - auto TODAY()
        ws.cell(row=r, column=26, value="=TODAY()")
        ws.cell(row=r, column=26).font = Font(color="0000FF")
        ws.cell(row=r, column=26).fill = styles["formula_cell"]["fill"]

    ws.freeze_panes = "A37"


# ============================================================================
# SECTION 6: KPIs & METRICS
# ============================================================================

def create_kpis_metrics(ws, styles):
    """Create comprehensive KPIs & Metrics tracking sheet."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "KEY PERFORMANCE INDICATORS & METRICS\n"
        "Secure Coding program effectiveness measurement"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- KPI CATEGORY 1: COMPLIANCE METRICS ----------
    row = 3
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "COMPLIANCE METRICS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    kpi_headers = ["KPI", "Current Value", "Target", "Threshold", "Status", "Trend", "Last Quarter", "YoY Change"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    compliance_kpis = [
        ("Overall Secure Coding Compliance Rate", "='Executive Dashboard'!C29", "≥95%", "90%"),
        ("Data Transmission Encryption Coverage", "", "100%", "95%"),
        ("Data at Rest Encryption Coverage", "", "100%", "95%"),
        ("Strong Authentication Coverage", "", "≥90%", "80%"),
        ("Key Management Compliance", "", "≥95%", "90%"),
        ("Algorithm Compliance Rate", "", "100%", "98%"),
        ("Certificate Validity Compliance", "", "100%", "95%"),
        ("Password Hashing Compliance", "", "100%", "100%"),
        ("Service Account Security", "", "≥90%", "85%"),
        ("SSO Coverage", "", "≥80%", "70%"),
        ("MFA Enrollment - All Users", "", "100%", "95%"),
        ("MFA Enrollment - Administrators", "", "100%", "100%"),
        ("Policy Compliance Score", "", "≥95%", "90%"),
        ("Audit Findings - Open", "", "0", "≤3"),
        ("Compliance Exceptions", "", "≤5", "≤10"),
    ]

    row += 1
    for kpi_name, current, target, threshold in compliance_kpis:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=2, value=current)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=threshold)
        
        if "=" in str(current):
            ws.cell(row=row, column=2).font = Font(color="0000FF")
        else:
            ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        
        for col in [5, 6, 7, 8]:  # Status, Trend, Last Quarter, YoY Change
            ws.cell(row=row, column=col).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles["border"]
        
        row += 1

    # ---------- KPI CATEGORY 2: RISK METRICS ----------
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "RISK METRICS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    risk_kpis = [
        ("Critical Risks - Open", '=COUNTIFS(\'Risk Register\'!N20:N119,">=20",\'Risk Register\'!Q20:Q119,"<>Closed")', "0", "0"),
        ("High Risks - Open", '=COUNTIFS(\'Risk Register\'!N20:N119,">=15",\'Risk Register\'!Q20:Q119,"<>Closed")', "≤5", "≤10"),
        ("Average Risk Score", '=AVERAGE(\'Risk Register\'!N20:N119)', "<10", "<15"),
        ("Risk Reduction Rate", "", "≥20% per Q", "≥10%"),
        ("Residual Risk Exposure", '=SUM(\'Risk Register\'!N20:N119)', "<100", "<200"),
        ("Risks Accepted (No Remediation)", '=COUNTIF(\'Risk Register\'!Q20:Q119,"Accepted")', "≤5", "≤10"),
        ("Risks Transferred", "", "-", "-"),
        ("Time to Risk Identification (Avg Days)", "", "≤7", "≤14"),
        ("Time to Risk Remediation (Avg Days)", "", "≤90", "≤120"),
        ("Overdue Risks", "", "0", "≤2"),
    ]

    row += 1
    for kpi_name, current, target, threshold in risk_kpis:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=2, value=current)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=threshold)
        
        if "=" in str(current):
            ws.cell(row=row, column=2).font = Font(color="0000FF")
        else:
            ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        
        for col in [5, 6, 7, 8]:
            ws.cell(row=row, column=col).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles["border"]
        
        row += 1

    # ---------- KPI CATEGORY 3: REMEDIATION METRICS ----------
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "REMEDIATION METRICS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    remediation_kpis = [
        ("Total Open Gaps", '=COUNTIF(\'Gap Analysis\'!R12:R211,"<>Closed")', "≤10", "≤20"),
        ("Critical Gaps - Open", '=COUNTIFS(\'Gap Analysis\'!I12:I211,"Critical",\'Gap Analysis\'!R12:R211,"<>Closed")', "0", "0"),
        ("High Priority Gaps - Open", '=COUNTIFS(\'Gap Analysis\'!I12:I211,"High",\'Gap Analysis\'!R12:R211,"<>Closed")', "≤5", "≤10"),
        ("Gaps Closed This Quarter", "", "≥10", "≥5"),
        ("Gap Closure Rate", '=COUNTIF(\'Gap Analysis\'!R12:R211,"Closed")/COUNTA(\'Gap Analysis\'!R12:R211)*100&"%"', "≥80%", "≥70%"),
        ("Remediation On-Time %", "", "≥90%", "≥80%"),
        ("Overdue Remediation Items", '=COUNTIFS(\'Remediation Roadmap\'!M37:M236,"<"&TODAY(),\'Remediation Roadmap\'!H37:H236,"<>Completed")', "0", "≤3"),
        ("Average Time to Remediation (Days)", "", "≤90", "≤120"),
        ("Budget Utilization Rate", "", "90-100%", "80-110%"),
        ("Remediation Success Rate", "", "≥95%", "≥90%"),
        ("Recurring Gaps", "", "0", "≤2"),
    ]

    row += 1
    for kpi_name, current, target, threshold in remediation_kpis:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=2, value=current)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=threshold)
        
        if "=" in str(current):
            ws.cell(row=row, column=2).font = Font(color="0000FF")
        else:
            ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        
        for col in [5, 6, 7, 8]:
            ws.cell(row=row, column=col).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles["border"]
        
        row += 1

    # ---------- KPI CATEGORY 4: TECHNICAL METRICS ----------
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "TECHNICAL IMPLEMENTATION METRICS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    technical_kpis = [
        ("Systems Using Approved Algorithms", "", "100%", "98%"),
        ("Systems Using Weak/Legacy Crypto", "", "0", "≤2"),
        ("TLS 1.3 Adoption Rate", "", "≥80%", "≥60%"),
        ("TLS 1.2+ Coverage", "", "100%", "100%"),
        ("Certificate Auto-Renewal Rate", "", "≥90%", "≥75%"),
        ("Certificates Expiring <30 Days", "", "0", "≤5"),
        ("Expired Certificates", "", "0", "0"),
        ("Self-Signed Certs in Production", "", "0", "0"),
        ("Key Rotation Compliance", "", "≥95%", "≥90%"),
        ("Keys Stored in HSM/KMS", "", "≥90%", "≥80%"),
        ("FIPS 140-2 Validated Modules", "", "≥80%", "≥70%"),
        ("Database Encryption (TDE) Coverage", "", "100%", "95%"),
        ("Disk Encryption Coverage - Endpoints", "", "100%", "98%"),
        ("Disk Encryption Coverage - Servers", "", "100%", "95%"),
        ("Backup Encryption Coverage", "", "100%", "100%"),
    ]

    row += 1
    for kpi_name, current, target, threshold in technical_kpis:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=2, value=current)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=threshold)
        
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        
        for col in [5, 6, 7, 8]:
            ws.cell(row=row, column=col).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles["border"]
        
        row += 1

    # Column widths
    widths = [35, 16, 14, 14, 12, 12, 14, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A4"


# ============================================================================
# END OF PART 2C
# ============================================================================

# ============================================================================
# SECTION 7: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create consolidated evidence register for audit traceability."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:V1")
    ws["A1"] = (
        "CONSOLIDATED EVIDENCE REGISTER\n"
        "Central repository for all cryptography compliance evidence"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- EVIDENCE SUMMARY ----------
    ws["A3"] = "Evidence Summary Statistics"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    summary_headers = [
        "Assessment Area",
        "Total Evidence Items",
        "Verified",
        "Pending Verification",
        "Not Verified",
        "Verification %"
    ]
    
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = ["Data Transmission", "Data Storage", "Authentication", "Key Management"]
    
    row += 1
    summary_start = row
    for area in areas:
        ws.cell(row=row, column=1, value=area)
        ws.cell(row=row, column=2, value=f'=COUNTIF(B17:B516,"{area}")')
        ws.cell(row=row, column=3, value=f'=COUNTIFS(B17:B516,"{area}",K17:K516,"Verified")')
        ws.cell(row=row, column=4, value=f'=COUNTIFS(B17:B516,"{area}",K17:K516,"Pending Verification")')
        ws.cell(row=row, column=5, value=f'=COUNTIFS(B17:B516,"{area}",K17:K516,"Not Verified")')
        ws.cell(row=row, column=6, value=f'=IF(B{row}=0,"0%",C{row}/B{row}*100&"%")')
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = styles["border"]
            if col >= 2:
                ws.cell(row=row, column=col).font = Font(color="0000FF")
        row += 1

    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 7):
        ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{summary_start}:{get_column_letter(col)}{row-1})")
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = styles["border"]

    # Column widths for summary
    widths_summary = [22, 18, 14, 18, 14, 16]
    for col_idx, width in enumerate(widths_summary, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---------- DETAILED EVIDENCE REGISTER ----------
    row += 3
    ws.merge_cells(f"A{row}:V{row}")
    ws[f"A{row}"] = "DETAILED EVIDENCE REGISTER"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    detail_headers = [
        "Evidence ID",
        "Assessment Area",
        "Source Document",
        "Gap ID (Link)",
        "Control Reference",
        "Evidence Type",
        "Evidence Description",
        "Evidence Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
        "Verified By",
        "Verification Date",
        "Retention Period",
        "Destruction Date",
        "Classification",
        "Accessibility",
        "Related Risk ID",
        "Related Remediation ID",
        "Audit Reference",
        "Notes/Comments",
        "Last Updated"
    ]
    
    widths_detail = [14, 22, 22, 12, 20, 22, 40, 45, 14, 20, 18, 20, 14, 16, 14, 16, 18, 14, 16, 18, 30, 14]
    
    for col_idx, (header, width) in enumerate(zip(detail_headers, widths_detail), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Create validations
    dv_area = DataValidation(
        type="list",
        formula1='"Data Transmission,Data Storage,Authentication,Key Management"',
        allow_blank=False
    )
    ws.add_data_validation(dv_area)

    dv_source = DataValidation(
        type="list",
        formula1='"ISMS-IMP-A.8.28.1,ISMS-IMP-A.8.28.2,ISMS-IMP-A.8.28.3,ISMS-IMP-A.8.28.4"',
        allow_blank=False
    )
    ws.add_data_validation(dv_source)

    dv_evidence_type = DataValidation(
        type="list",
        formula1='"Configuration File,Screenshot,Policy Document,Audit Log,Test Result,Scan Report,Certificate,Procedure Document,Training Record,Meeting Minutes,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_evidence_type)

    dv_verification = DataValidation(
        type="list",
        formula1='"Verified,Pending Verification,Not Verified,Requires Update,Expired"',
        allow_blank=False
    )
    ws.add_data_validation(dv_verification)

    dv_retention = DataValidation(
        type="list",
        formula1='"1 year,3 years,5 years,7 years,Permanent"',
        allow_blank=False
    )
    ws.add_data_validation(dv_retention)

    dv_classification = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=False
    )
    ws.add_data_validation(dv_classification)

    dv_accessibility = DataValidation(
        type="list",
        formula1='"Shared Drive,SharePoint,Document Management System,Physical Storage,Cloud Storage"',
        allow_blank=False
    )
    ws.add_data_validation(dv_accessibility)

    # Data entry rows (500 rows for evidence)
    row += 1
    data_start = row
    
    for r in range(row, row + 500):
        # Evidence ID auto-increment
        ws.cell(row=r, column=1, value=f"EV-{r-row+1:03d}").font = Font(color="808080")
        
        # Input cells
        for col in range(2, 23):
            cell = ws.cell(row=r, column=col)
            
            # Skip formula columns
            if col not in [15, 22]:  # Destruction Date, Last Updated
                cell.fill = styles["input_cell"]["fill"]
            
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        
        # Apply validations
        dv_area.add(ws.cell(row=r, column=2))           # B: Assessment Area
        dv_source.add(ws.cell(row=r, column=3))         # C: Source Document
        dv_evidence_type.add(ws.cell(row=r, column=6))  # F: Evidence Type
        dv_verification.add(ws.cell(row=r, column=11))  # K: Verification Status
        dv_retention.add(ws.cell(row=r, column=14))     # N: Retention Period
        dv_classification.add(ws.cell(row=r, column=16))# P: Classification
        dv_accessibility.add(ws.cell(row=r, column=17)) # Q: Accessibility
        
        # Destruction Date (column O = 15) - auto-calculate from collection date + retention
        ws.cell(row=r, column=15, value=f'=IF(OR(ISBLANK(I{r}),ISBLANK(N{r})),"",EDATE(I{r},IF(N{r}="1 year",12,IF(N{r}="3 years",36,IF(N{r}="5 years",60,IF(N{r}="7 years",84,600))))))')
        ws.cell(row=r, column=15).font = Font(color="0000FF")
        ws.cell(row=r, column=15).fill = styles["formula_cell"]["fill"]
        
        # Last Updated (column V = 22) - auto TODAY()
        ws.cell(row=r, column=22, value="=TODAY()")
        ws.cell(row=r, column=22).font = Font(color="0000FF")
        ws.cell(row=r, column=22).fill = styles["formula_cell"]["fill"]

    ws.freeze_panes = "A17"


# ============================================================================
# SECTION 8: ACTION ITEMS & FOLLOW-UP
# ============================================================================

def create_action_items(ws, styles):
    """Create action items and follow-up tracker."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:X1")
    ws["A1"] = (
        "ACTION ITEMS & FOLLOW-UP TRACKER\n"
        "Centralized tracking of all outstanding actions"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- ACTION SUMMARY DASHBOARD ----------
    ws["A3"] = "Action Summary Dashboard"
    ws["A3"].font = Font(bold=True, size=11)

    # Current Status
    row = 4
    ws[f"A{row}"] = "Current Status"
    ws[f"A{row}"].font = Font(bold=True)
    
    row += 1
    status_headers = ["Status", "Count", "Percentage", "Overdue Count"]
    for col_idx, header in enumerate(status_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    statuses = ["Not Started", "In Progress", "Blocked", "Pending Review", "Completed"]
    row += 1
    status_start = row
    
    for status in statuses:
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=f'=COUNTIF(G22:G221,"{status}")')
        ws.cell(row=row, column=3, value=f'=IF(B{status_start+4}=0,"0%",B{row}/B{status_start+4}*100&"%")')
        if status != "Completed":
            ws.cell(row=row, column=4, value=f'=COUNTIFS(G22:G221,"{status}",K22:K221,"<"&TODAY())')
        else:
            ws.cell(row=row, column=4, value="N/A")
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = styles["border"]
            if col >= 2 and status != "Completed":
                ws.cell(row=row, column=col).font = Font(color="0000FF")
        row += 1

    # TOTAL
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{status_start}:B{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"=SUM(D{status_start}:D{row-1})").font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = styles["border"]

    # By Priority
    row += 2
    ws[f"A{row}"] = "By Priority"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    priority_headers = ["Priority", "Open Actions", "Overdue", "% Complete", "Avg Days Open"]
    for col_idx, header in enumerate(priority_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    priorities = ["Critical", "High", "Medium", "Low"]
    row += 1
    
    for priority in priorities:
        ws.cell(row=row, column=1, value=priority)
        ws.cell(row=row, column=2, value=f'=COUNTIFS(F22:F221,"{priority}",G22:G221,"<>Completed")')
        ws.cell(row=row, column=3, value=f'=COUNTIFS(F22:F221,"{priority}",K22:K221,"<"&TODAY(),G22:G221,"<>Completed")')
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = styles["border"]
            if col in [2, 3]:
                ws.cell(row=row, column=col).font = Font(color="0000FF")
        row += 1

    # Column widths
    widths_summary = [16, 14, 14, 14, 16]
    for col_idx, width in enumerate(widths_summary, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---------- DETAILED ACTION REGISTER ----------
    row += 2
    ws.merge_cells(f"A{row}:X{row}")
    ws[f"A{row}"] = "DETAILED ACTION REGISTER"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    detail_headers = [
        "Action ID",
        "Source",
        "Source ID",
        "Action Description",
        "Action Type",
        "Priority",
        "Status",
        "Owner",
        "Supporting Team",
        "Created Date",
        "Due Date",
        "Completed Date",
        "Days Open",
        "Days Remaining",
        "On Time Status",
        "Blocking Issue",
        "Dependencies",
        "Progress %",
        "Last Update Date",
        "Last Update By",
        "Update Notes",
        "Escalation Required",
        "Escalated To",
        "Escalation Date"
    ]
    
    widths_detail = [12, 20, 14, 40, 20, 12, 16, 20, 20, 14, 14, 14, 10, 10, 14, 30, 30, 10, 14, 20, 35, 14, 20, 14]
    
    for col_idx, (header, width) in enumerate(zip(detail_headers, widths_detail), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Create validations
    dv_source = DataValidation(
        type="list",
        formula1='"Gap Analysis,Risk Register,Audit Finding,Management Review,Incident,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_source)

    dv_action_type = DataValidation(
        type="list",
        formula1='"Investigation,Remediation,Documentation,Training,Process Change,Tool Implementation,Review,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_action_type)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Blocked,Pending Review,Completed,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    # Data entry rows (200 rows for actions)
    row += 1
    data_start = row
    
    for r in range(row, row + 200):
        # Action ID auto-increment
        ws.cell(row=r, column=1, value=f"ACT-{r-row+1:03d}").font = Font(color="808080")
        
        # Input cells
        for col in range(2, 25):
            cell = ws.cell(row=r, column=col)
            
            # Skip formula columns
            if col not in [13, 14, 15]:  # Days Open, Days Remaining, On Time Status
                cell.fill = styles["input_cell"]["fill"]
            
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        
        # Apply validations
        dv_source.add(ws.cell(row=r, column=2))        # B: Source
        dv_action_type.add(ws.cell(row=r, column=5))   # E: Action Type
        dv_priority.add(ws.cell(row=r, column=6))      # F: Priority
        dv_status.add(ws.cell(row=r, column=7))        # G: Status
        dv_yesno.add(ws.cell(row=r, column=22))        # V: Escalation Required
        
        # Days Open (column M = 13)
        ws.cell(row=r, column=13, value=f'=IF(G{r}="Completed",IF(ISBLANK(L{r}),"",L{r}-J{r}),IF(ISBLANK(J{r}),"",TODAY()-J{r}))')
        ws.cell(row=r, column=13).font = Font(color="0000FF")
        ws.cell(row=r, column=13).fill = styles["formula_cell"]["fill"]
        
        # Days Remaining (column N = 14)
        ws.cell(row=r, column=14, value=f'=IF(G{r}="Completed",0,IF(ISBLANK(K{r}),"",K{r}-TODAY()))')
        ws.cell(row=r, column=14).font = Font(color="0000FF")
        ws.cell(row=r, column=14).fill = styles["formula_cell"]["fill"]
        
        # On Time Status (column O = 15)
        ws.cell(row=r, column=15, value=f'=IF(G{r}="Completed","Completed",IF(ISBLANK(N{r}),"",IF(N{r}<0,"Overdue",IF(N{r}<7,"At Risk","On Time"))))')
        ws.cell(row=r, column=15).font = Font(color="0000FF")
        ws.cell(row=r, column=15).fill = styles["formula_cell"]["fill"]
        
        # Progress % format
        ws.cell(row=r, column=18).number_format = '0"%"'

    ws.freeze_panes = "A22"


# ============================================================================
# SECTION 9: AUDIT & COMPLIANCE LOG
# ============================================================================

def create_audit_log(ws, styles):
    """Create audit and compliance log."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:T1")
    ws["A1"] = (
        "AUDIT & COMPLIANCE LOG\n"
        "Record of all audits, assessments, and compliance reviews"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- AUDIT SUMMARY ----------
    ws["A3"] = "Audit Summary"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    summary_headers = [
        "Audit Type",
        "Total Audits",
        "Open Findings",
        "Closed Findings",
        "Avg Time to Close",
        "Compliance Rate"
    ]
    
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    audit_types = [
        "Internal Audit",
        "External Audit",
        "Self-Assessment",
        "Vendor Audit",
        "Regulatory Inspection"
    ]
    
    row += 1
    summary_start = row
    for audit_type in audit_types:
        ws.cell(row=row, column=1, value=audit_type)
        ws.cell(row=row, column=2, value=f'=COUNTIF(B19:B118,"{audit_type}")')
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = styles["border"]
        if row > summary_start:
            ws.cell(row=row, column=2).font = Font(color="0000FF")
        row += 1

    # TOTAL
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{summary_start}:B{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(C{summary_start}:C{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"=SUM(D{summary_start}:D{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"=AVERAGE(E{summary_start}:E{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=AVERAGE(F{summary_start}:F{row-1})").font = Font(bold=True)
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = styles["border"]

    # Column widths
    widths_summary = [22, 14, 14, 14, 16, 16]
    for col_idx, width in enumerate(widths_summary, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---------- AUDIT REGISTER ----------
    row += 3
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "AUDIT REGISTER"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    audit_headers = [
        "Audit ID",
        "Audit Type",
        "Audit Scope",
        "Assessment Area",
        "Auditor/Assessor",
        "Audit Date",
        "Audit Report Date",
        "Total Findings",
        "Critical Findings",
        "High Findings",
        "Medium Findings",
        "Low Findings",
        "Open Findings",
        "Closed Findings",
        "Compliance Score",
        "Report Location",
        "Follow-up Required",
        "Follow-up Date",
        "Status",
        "Notes"
    ]
    
    widths_audit = [12, 20, 30, 22, 25, 14, 14, 12, 12, 12, 12, 12, 12, 12, 14, 35, 14, 14, 16, 30]
    
    for col_idx, (header, width) in enumerate(zip(audit_headers, widths_audit), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Create validations
    dv_audit_type = DataValidation(
        type="list",
        formula1='"Internal,External,Self-Assessment,Vendor,Regulatory,Penetration Test,Compliance Scan"',
        allow_blank=False
    )
    ws.add_data_validation(dv_audit_type)

    dv_area = DataValidation(
        type="list",
        formula1='"All,Data Transmission,Data Storage,Authentication,Key Management"',
        allow_blank=False
    )
    ws.add_data_validation(dv_area)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    dv_status = DataValidation(
        type="list",
        formula1='"Scheduled,In Progress,Complete,Follow-up Required"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Data entry rows (100 rows)
    row += 1
    
    for r in range(row, row + 100):
        ws.cell(row=r, column=1, value=f"AUD-{r-row+1:03d}").font = Font(color="808080")
        
        for col in range(2, 21):
            cell = ws.cell(row=r, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        
        dv_audit_type.add(ws.cell(row=r, column=2))
        dv_area.add(ws.cell(row=r, column=4))
        dv_yesno.add(ws.cell(row=r, column=17))
        dv_status.add(ws.cell(row=r, column=19))
        
        # Compliance Score as percentage
        ws.cell(row=r, column=15).number_format = '0"%"'

    ws.freeze_panes = "A19"


# ============================================================================
# SECTION 10: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create approval and sign-off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Dashboard Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.8.28.5 - Compliance Summary Dashboard"),
        ("Assessment Period", ""),
        ("Portfolio Compliance (Average)", "='Executive Dashboard'!C29"),
        ("Dashboard Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        
        if "=" in str(value):
            ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        else:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Status dropdown
    status_cell_row = row - 1
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{status_cell_row}"])

    # Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # ---------- REVIEWED BY (INFORMATION SECURITY OFFICER) ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (INFORMATION SECURITY OFFICER)"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    review_fields = ["Name", "Date", "Review Notes"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # ---------- APPROVED BY (CISO) ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Conditions/Notes"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approval Decision dropdown (3rd field from approval section start)
    decision_cell_row = row - 2
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{decision_cell_row}"])

    # ---------- NEXT REVIEW DETAILS ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    next_review_fields = ["Next Review Date", "Review Responsible", "Special Considerations"]
    row += 1
    for field in next_review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 11: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.28.5 - Compliance Summary Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.28: Secure Coding")
    logger.info("=" * 80)

    wb = create_workbook()
    styles = setup_styles()

    logger.info("\n[1/9] Creating Executive Dashboard sheet...")
    create_executive_dashboard(wb["Executive Dashboard"], styles)

    logger.info("[2/9] Creating Gap Analysis sheet...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[3/9] Creating Risk Register sheet...")
    create_risk_register(wb["Risk Register"], styles)

    logger.info("[4/9] Creating Remediation Roadmap sheet...")
    create_remediation_roadmap(wb["Remediation Roadmap"], styles)

    logger.info("[5/9] Creating KPIs & Metrics sheet...")
    create_kpis_metrics(wb["KPIs & Metrics"], styles)

    logger.info("[6/9] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[7/9] Creating Action Items & Follow-up sheet...")
    create_action_items(wb["Action Items & Follow-up"], styles)

    logger.info("[8/9] Creating Audit & Compliance Log sheet...")
    create_audit_log(wb["Audit & Compliance Log"], styles)

    logger.info("[9/9] Creating Approval Sign-Off sheet...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.8.28.5_Compliance_Summary_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\n" + "=" * 80)
    logger.info("COMPLIANCE SUMMARY DASHBOARD - COMPLETE")
    logger.info("=" * 80)
    logger.info("\nWorkbook Structure (9 Sheets):")
    logger.info("  1. Executive Dashboard - Overall compliance status with external links")
    logger.info("  2. Gap Analysis - 200 gap entries with risk scoring")
    logger.info("  3. Risk Register - 100 risk entries with inherent/residual scoring")
    logger.info("  4. Remediation Roadmap - 200 remediation items with timeline tracking")
    logger.info("  5. KPIs & Metrics - 50+ KPIs across 4 categories")
    logger.info("  6. Evidence Register - 500 evidence entries with retention tracking")
    logger.info("  7. Action Items & Follow-up - 200 action items with status tracking")
    logger.info("  8. Audit & Compliance Log - 100 audit records")
    logger.info("  9. Approval Sign-Off - Multi-level approval workflow")
    
    logger.info("\nExternal Links Configuration:")
    logger.info("  \u2022 Dashboard references 4 normalized source workbooks:")
    logger.info("    - ISMS-IMP-A.8.28.1.xlsx (SDLC Assessment)")
    logger.info("    - ISMS-IMP-A.8.28.2.xlsx (Standards & Tools Assessment)")
    logger.info("    - ISMS-IMP-A.8.28.3.xlsx (Code Review & Testing Assessment)")
    logger.info("    - ISMS-IMP-A.8.28.4.xlsx (Third-Party & OSS Assessment)")
    
    logger.info("\nIMPORTANT SETUP STEPS:")
    logger.info("  1. Run normalization script FIRST:")
    logger.info("     python3 normalize_assessment_files_a828.py")
    logger.info("")
    logger.info("  2. Place this dashboard in the Dashboard_Sources folder")
    logger.info("     alongside the 4 normalized source workbooks")
    logger.info("")
    logger.info("  3. Open dashboard and click 'Update Links' when prompted")
    logger.info("")
    logger.info("  4. Executive Dashboard will auto-populate with compliance data!")
    logger.info("")
    logger.info("  5. Complete manual entry sections:")
    logger.info("     - Gap Analysis (consolidated gaps)")
    logger.info("     - Risk Register (identified risks)")
    logger.info("     - Remediation Roadmap (action plans)")
    logger.info("     - KPIs (current metric values)")
    logger.info("     - Evidence Register (supporting documentation)")
    logger.info("")
    logger.info("=" * 80)
    logger.info("\n🎄 ALL 5 SECURE CODING ASSESSMENT TOOLS COMPLETE! 🎄")
    logger.info("\nYou now have:")
    logger.info("  \u2705 4 Assessment workbooks (SDLC, Standards & Tools, Code Review & Testing, Third-Party & OSS)")
    logger.info("  \u2705 1 Normalization script (prepares files for dashboard)")
    logger.info("  \u2705 1 Consolidated dashboard (executive oversight)")
    logger.info("\nMerry Christmas, Greg! Now go enjoy your holiday! 🎅🎁")
    logger.info("=" * 80 + "\n")


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT - COMPLETE COMPLIANCE SUMMARY DASHBOARD GENERATOR
# ============================================================================

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
