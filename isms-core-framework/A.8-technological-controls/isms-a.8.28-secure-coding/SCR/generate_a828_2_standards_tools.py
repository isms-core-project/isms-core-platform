#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.28.2 - Standards & Tools Assessment
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Assessment Domain 2 of 4: Coding Standards and Development Tool Security

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific programming languages, frameworks, development
tools, and coding standard requirements.

Key customisation areas:
1. Programming languages and frameworks (match your technology stack)
2. Coding standards and style guides (specific to your language choices)
3. Development tools and IDE configurations (based on your toolchain)
4. Security linting rules and enforcement (adapt to your risk profile)
5. Compliance criteria and scoring (aligned with your quality requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMISE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
secure coding standards implementation and development tool security configuration
across all programming languages and platforms in use.

**Purpose:**
Enables systematic assessment of coding standards and tool security against
ISO 27001:2022 Control A.8.28 requirements, supporting evidence-based validation
of secure development practices and tool-enforced security controls.

**Assessment Scope:**
- Language-specific secure coding standards (OWASP, CWE, CERT, vendor guides)
- IDE and editor security plugins and configurations
- Static Application Security Testing (SAST) tool implementation
- Security linting rules and enforcement mechanisms
- Code formatting and style guide compliance
- Dependency management and vulnerability scanning tools
- Secret detection and credential management
- Pre-commit hooks and automated security checks
- Developer workstation security hardening
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and coding standard references
2. Language Standards - Per-language secure coding standard implementation
3. SAST Tools - Static analysis tool coverage and configuration
4. IDE Security - Development environment security plugin assessment
5. Linting & Formatting - Security linter rules and enforcement
6. Dependency Management - SCA tools and vulnerability scanning
7. Secret Detection - Credential and API key protection mechanisms
8. Pre-Commit Controls - Automated security gate implementation
9. Workstation Security - Developer endpoint hardening assessment
10. Gap Analysis - Missing or inadequate standards/tools and remediation
11. Evidence Register - Audit evidence tracking and documentation
12. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with language and tool dropdown lists
- Conditional formatting for standard compliance and tool coverage
- Automated gap identification for unsupported languages or missing tools
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with SAST/SCA tool configuration exports

**Integration:**
This assessment feeds into the A.8.28.5 Compliance Dashboard, which
consolidates data from all four assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a828_2_standards_tools.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a828_2_standards_tools.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a828_2_standards_tools.py --date 20250124

Output:
    File: ISMS_A_8_28_2_Standards_Tools_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise language-specific coding standards to match tech stack
    2. Inventory all programming languages, frameworks, and tools in use
    3. Complete assessments for each language and development tool
    4. Validate SAST/SCA tool coverage and rule configuration
    5. Review pre-commit hook implementation and enforcement
    6. Conduct gap analysis for languages without standards or inadequate tooling
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (tool configs, scan results, standard docs)
    9. Obtain stakeholder approvals (Dev Leads, AppSec, Engineering Manager)
    10. Feed results into A.8.28.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    2 of 4 (Standards & Tools)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.2: Standards & Tools Implementation Guide
    - ISMS-IMP-A.8.28.1: SDLC Integration Assessment (Domain 1)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Assessment (Domain 3)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Assessment (Domain 4)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a828_1_sdlc_assessment.py
    - generate_a828_3_code_review_testing.py
    - generate_a828_4_third_party_oss.py
    - generate_a828_5_compliance_dashboard.py
    - normalize_assessment_files_a828.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.28.2 specification
    - Supports comprehensive coding standards and tool security evaluation
    - Integrated with A.8.28.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Technology Stack Specificity:**
This assessment framework is technology-agnostic by design. Customise language
lists, coding standards, and tool recommendations to match your organisation's
specific programming languages, frameworks, and development ecosystem.

**Tool Selection Philosophy:**
Prioritize tools that integrate seamlessly with existing workflows over
"best-of-breed" tools that developers won't use. An adequate tool that's
actually used is better than a perfect tool that's bypassed.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of SAST/SCA tool configurations, linting
rules, and coding standard adoption.

**Data Protection:**
Assessment workbooks contain sensitive development information including:
- Technology stack details and framework versions
- Security tool configurations and rule sets
- Vulnerability patterns and weakness trends
- Development environment details

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check SAST/SCA rule updates and new language support
- Semi-annually: Review coding standard versions and update references
- Annually: Complete reassessment of all languages and tools in use
- Ad-hoc: When new languages/frameworks adopted or tools changed

**Quality Assurance:**
Have application security SMEs, technical leads for each language/platform,
and DevOps engineers validate assessments before using results for compliance
reporting or remediation decisions.

**Regulatory Alignment:**
Ensure coding standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 secure coding standards (Requirement 6.2)
- Healthcare: HIPAA secure development practices
- Finance: Regional banking secure coding requirements
- Government: NIST Secure Software Development Framework (SSDF)

Customise assessment criteria to include regulatory-specific requirements.

**Integration with Development Tools:**
Where possible, integrate assessment data with:
- SAST tool outputs (SonarQube, Checkmarx, Fortify, Semgrep, etc.)
- SCA tool results (Snyk, WhiteSource, Black Duck, Dependabot, etc.)
- IDE plugin configurations (export from VS Code, IntelliJ, etc.)
- Pre-commit hook repositories (Git hook configurations)

Automated tool integration reduces manual effort and improves accuracy.

**Don't Fool Yourself - Feynman Principle:**
Having a coding standard document doesn't mean developers follow it. Having
SAST tools doesn't mean they're configured correctly or that findings are
addressed. Assess actual tool usage and standard adoption:
- Are SAST scans running on every commit/build?
- What percentage of findings are actually fixed vs. marked "Won't Fix"?
- Can you show code that was rejected for standard violations?
- Are developers actually using IDE security plugins?

Evidence of use > Existence of tools.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.28.2"
WORKBOOK_NAME = "Coding Standards and Development Tool Security"
CONTROL_ID = "A.8.28"
CONTROL_NAME = "Secure Coding"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


import sys

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.28.2 specification
    # 10 sheets total: Instructions + 5 assessment domains + 4 supporting sheets
    sheets = [
        "Instructions & Legend",
        "Coding Standards Adoption",
        "SAST SCA Tools",
        "DAST Security Testing Tools",
        "IDE Plugins Linters",
        "Tool Effectiveness Metrics",
        "Evidence Register",
        "Gap Analysis",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Returns style templates as dictionaries to avoid Excel's shared object warnings.
    Each cell gets a fresh style instance - no cargo cult object sharing.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_implemented": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notimplemented": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        },
        "priority_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "priority_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "priority_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "priority_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    
    Tools are only as good as their configuration and usage.
    These validations help assess deployment maturity, not just presence.
    """
    validations = {
        # Implementation status dropdowns
        'implementation_status': DataValidation(
            type="list",
            formula1='"✅ Implemented,⚠️ Partially Implemented,❌ Not Implemented,N/A"',
            allow_blank=False
        ),
        'deployment_status': DataValidation(
            type="list",
            formula1='"Deployed,Pilot,Planned,Not Deployed"',
            allow_blank=False
        ),
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Unknown"',
            allow_blank=False
        ),
        
        # Tool maturity and effectiveness
        'maturity_level': DataValidation(
            type="list",
            formula1='"Mature,Operational,Developing,Initial,Not Deployed"',
            allow_blank=False
        ),
        'effectiveness_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Fair,Poor,Not Assessed"',
            allow_blank=False
        ),
        'adoption_rate': DataValidation(
            type="list",
            formula1='"90-100%,70-89%,50-69%,Below 50%,Unknown"',
            allow_blank=False
        ),
        'coverage': DataValidation(
            type="list",
            formula1='"Complete,High (>75%),Medium (50-75%),Low (<50%),None"',
            allow_blank=False
        ),
        
        # Priority and risk levels
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Negligible"',
            allow_blank=False
        ),
        
        # Tool integration
        'integration_level': DataValidation(
            type="list",
            formula1='"Fully Integrated,Partially Integrated,Standalone,Not Integrated"',
            allow_blank=False
        ),
        'automation_level': DataValidation(
            type="list",
            formula1='"Fully Automated,Partially Automated,Manual,Not Applicable"',
            allow_blank=False
        ),
        
        # Gap analysis
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Closed,Deferred"',
            allow_blank=False
        ),
        
        # Evidence types
        'evidence_type': DataValidation(
            type="list",
            formula1='"Document,Screenshot,Report,Configuration,URL,Tool Output,Dashboard,Policy,Other"',
            allow_blank=False
        ),
        
        # Tool types
        'tool_type_sast': DataValidation(
            type="list",
            formula1='"SonarQube,Semgrep,Checkmarx,Fortify,Veracode,CodeQL,Other,None"',
            allow_blank=False
        ),
        'tool_type_sca': DataValidation(
            type="list",
            formula1='"Snyk,WhiteSource,Dependabot,Black Duck,OWASP Dependency-Check,Other,None"',
            allow_blank=False
        ),
        'tool_type_dast': DataValidation(
            type="list",
            formula1='"OWASP ZAP,Burp Suite,Acunetix,Netsparker,AppScan,Other,None"',
            allow_blank=False
        ),
        'tool_type_secrets': DataValidation(
            type="list",
            formula1='"TruffleHog,GitGuardian,git-secrets,detect-secrets,Gitleaks,Other,None"',
            allow_blank=False
        ),
        
        # Scan frequency
        'scan_frequency': DataValidation(
            type="list",
            formula1='"Every Commit,Daily,Weekly,On-Demand,Never"',
            allow_blank=False
        ),
        
        # Standards compliance
        'standards_compliance': DataValidation(
            type="list",
            formula1='"Fully Compliant,Mostly Compliant,Partially Compliant,Non-Compliant,Not Assessed"',
            allow_blank=False
        ),
        
        # Training status
        'training_status': DataValidation(
            type="list",
            formula1='"Completed,In Progress,Not Started,N/A"',
            allow_blank=False
        ),
        
        # Approval decision
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
            allow_blank=False
        ),
    }

    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: ASSESSMENT DOMAIN DATA DEFINITIONS
# ============================================================================

def get_domain1_requirements():
    """
    Domain 1: Coding Standards Adoption
    
    Standards without adoption are just documentation.
    This assesses whether secure coding standards are actually followed.
    """
    return [
        {
            "id": "1.1",
            "requirement": "Organisation has documented secure coding standards",
            "evidence": "Secure coding standards document, developer wiki, internal guidelines",
            "category": "Standards Documentation"
        },
        {
            "id": "1.2",
            "requirement": "Coding standards reference OWASP Top 10 and CWE Top 25",
            "evidence": "Standards document with OWASP/CWE mappings, policy references",
            "category": "Standards Documentation"
        },
        {
            "id": "1.3",
            "requirement": "Standards cover input validation requirements for all languages used",
            "evidence": "Language-specific input validation guidelines, code examples",
            "category": "Standards Documentation"
        },
        {
            "id": "1.4",
            "requirement": "Standards cover output encoding/escaping requirements",
            "evidence": "XSS prevention guidelines, encoding standards documentation",
            "category": "Standards Documentation"
        },
        {
            "id": "1.5",
            "requirement": "Standards cover authentication and session management requirements",
            "evidence": "Auth standards, session handling guidelines, JWT/token standards",
            "category": "Standards Documentation"
        },
        {
            "id": "1.6",
            "requirement": "Standards cover cryptography requirements (algorithms, key management)",
            "evidence": "Crypto standards, approved algorithms list, key management policies",
            "category": "Standards Documentation"
        },
        {
            "id": "1.7",
            "requirement": "Standards cover error handling and logging requirements",
            "evidence": "Logging standards, error handling patterns, log retention policies",
            "category": "Standards Documentation"
        },
        {
            "id": "1.8",
            "requirement": "Developers are trained on secure coding standards",
            "evidence": "Training completion records, training materials, attendance logs",
            "category": "Training & Awareness"
        },
        {
            "id": "1.9",
            "requirement": "Secure coding standards are easily accessible to all developers",
            "evidence": "Internal wiki links, IDE bookmarks, quick reference guides",
            "category": "Accessibility"
        },
        {
            "id": "1.10",
            "requirement": "Standards are reviewed and updated at least annually",
            "evidence": "Standards review schedule, change logs, version history",
            "category": "Standards Maintenance"
        },
        {
            "id": "1.11",
            "requirement": "Language-specific secure coding guides exist for primary languages",
            "evidence": "Python/Java/JavaScript/etc. specific security guidelines",
            "category": "Standards Documentation"
        },
        {
            "id": "1.12",
            "requirement": "Standards include code examples (secure vs. insecure patterns)",
            "evidence": "Code example repository, good/bad pattern documentation",
            "category": "Standards Documentation"
        },
        {
            "id": "1.13",
            "requirement": "Compliance with standards is measured and tracked",
            "evidence": "Compliance metrics, SAST rule compliance, audit results",
            "category": "Standards Enforcement"
        },
        {
            "id": "1.14",
            "requirement": "Standards violations are identified during code review",
            "evidence": "Code review checklists, review comments, violation tracking",
            "category": "Standards Enforcement"
        },
        {
            "id": "1.15",
            "requirement": "Developers can request clarifications or exceptions to standards",
            "evidence": "Exception request process, clarification tickets, security champion reviews",
            "category": "Standards Governance"
        },
        {
            "id": "1.16",
            "requirement": "Standards compliance is part of developer performance evaluation",
            "evidence": "Performance review criteria, security KPIs for developers",
            "category": "Standards Enforcement"
        },
        {
            "id": "1.17",
            "requirement": "Secure coding champions help developers apply standards",
            "evidence": "Security champion program documentation, champion activity logs",
            "category": "Training & Awareness"
        },
        {
            "id": "1.18",
            "requirement": "New hire onboarding includes secure coding standards training",
            "evidence": "Onboarding checklist, new hire training materials, completion records",
            "category": "Training & Awareness"
        },
    ]


def get_domain2_requirements():
    """
    Domain 2: SAST & SCA Tools
    
    Static analysis tools find vulnerabilities before they reach production.
    But only if they're configured correctly and findings are actually remediated.
    """
    return [
        {
            "id": "2.1",
            "requirement": "SAST (Static Application Security Testing) tool is deployed",
            "evidence": "Tool license, deployment documentation, access credentials",
            "category": "SAST Deployment"
        },
        {
            "id": "2.2",
            "requirement": "SAST tool covers all primary programming languages used",
            "evidence": "Language support documentation, configuration for each language",
            "category": "SAST Configuration"
        },
        {
            "id": "2.3",
            "requirement": "SAST scans run automatically on every code commit/PR",
            "evidence": "CI/CD integration config, pipeline logs showing SAST execution",
            "category": "SAST Integration"
        },
        {
            "id": "2.4",
            "requirement": "SAST tool is configured with security-focused rulesets",
            "evidence": "Ruleset configuration, enabled security rules list",
            "category": "SAST Configuration"
        },
        {
            "id": "2.5",
            "requirement": "SAST findings are prioritised by severity (Critical/High/Medium/Low)",
            "evidence": "Severity classification schema, findings dashboard",
            "category": "SAST Configuration"
        },
        {
            "id": "2.6",
            "requirement": "Critical/High SAST findings block code deployment",
            "evidence": "Quality gate configuration, blocked deployment logs",
            "category": "SAST Integration"
        },
        {
            "id": "2.7",
            "requirement": "SAST false positives can be suppressed with justification",
            "evidence": "Suppression process, suppression audit trail, justification records",
            "category": "SAST Configuration"
        },
        {
            "id": "2.8",
            "requirement": "SAST findings are tracked to resolution in ticketing system",
            "evidence": "JIRA/Azure DevOps integration, vulnerability tracking tickets",
            "category": "SAST Remediation"
        },
        {
            "id": "2.9",
            "requirement": "SCA (Software Composition Analysis) tool is deployed",
            "evidence": "Tool license, deployment documentation, access credentials",
            "category": "SCA Deployment"
        },
        {
            "id": "2.10",
            "requirement": "SCA tool scans all dependency manifests (package.json, pom.xml, etc.)",
            "evidence": "Supported manifest files list, scan coverage reports",
            "category": "SCA Configuration"
        },
        {
            "id": "2.11",
            "requirement": "SCA scans run automatically on every build",
            "evidence": "CI/CD integration, build logs showing SCA execution",
            "category": "SCA Integration"
        },
        {
            "id": "2.12",
            "requirement": "SCA tool checks for known vulnerabilities in dependencies (CVEs)",
            "evidence": "Vulnerability database updates, CVE detection reports",
            "category": "SCA Configuration"
        },
        {
            "id": "2.13",
            "requirement": "SCA tool checks for license compliance issues",
            "evidence": "License policy configuration, license violation reports",
            "category": "SCA Configuration"
        },
        {
            "id": "2.14",
            "requirement": "Critical/High vulnerability findings block deployment",
            "evidence": "Quality gate configuration, blocked deployment logs",
            "category": "SCA Integration"
        },
        {
            "id": "2.15",
            "requirement": "SCA tool provides remediation guidance (upgrade paths)",
            "evidence": "Remediation recommendations, automated PR generation",
            "category": "SCA Remediation"
        },
        {
            "id": "2.16",
            "requirement": "Dependency vulnerabilities have defined SLAs for remediation",
            "evidence": "Vulnerability remediation SLA policy, SLA tracking metrics",
            "category": "SCA Remediation"
        },
        {
            "id": "2.17",
            "requirement": "SAST and SCA findings are reviewed by security team regularly",
            "evidence": "Security review meeting notes, findings triage records",
            "category": "Tool Governance"
        },
        {
            "id": "2.18",
            "requirement": "Metrics track SAST/SCA finding trends and remediation velocity",
            "evidence": "Security metrics dashboard, trend analysis reports",
            "category": "Tool Effectiveness"
        },
    ]


def get_domain3_requirements():
    """
    Domain 3: DAST & Security Testing Tools
    
    Dynamic testing finds runtime vulnerabilities that static analysis misses.
    Testing in environments that mirror production is critical.
    """
    return [
        {
            "id": "3.1",
            "requirement": "DAST (Dynamic Application Security Testing) tool is deployed",
            "evidence": "Tool license, deployment documentation, scanning infrastructure",
            "category": "DAST Deployment"
        },
        {
            "id": "3.2",
            "requirement": "DAST scans run against staging/pre-production environments",
            "evidence": "Scan configuration, target environment documentation",
            "category": "DAST Configuration"
        },
        {
            "id": "3.3",
            "requirement": "DAST scans are executed before production releases",
            "evidence": "Release checklist with DAST requirement, scan reports",
            "category": "DAST Integration"
        },
        {
            "id": "3.4",
            "requirement": "DAST tool tests for OWASP Top 10 vulnerabilities",
            "evidence": "Scan policy configuration, OWASP Top 10 test coverage",
            "category": "DAST Configuration"
        },
        {
            "id": "3.5",
            "requirement": "DAST scans include authenticated testing (logged-in scenarios)",
            "evidence": "Authentication configuration, authenticated scan results",
            "category": "DAST Configuration"
        },
        {
            "id": "3.6",
            "requirement": "DAST findings are classified by severity and risk",
            "evidence": "Severity schema, risk scoring methodology",
            "category": "DAST Configuration"
        },
        {
            "id": "3.7",
            "requirement": "Critical DAST findings block production deployment",
            "evidence": "Release gate configuration, blocked release records",
            "category": "DAST Integration"
        },
        {
            "id": "3.8",
            "requirement": "DAST false positives are validated and documented",
            "evidence": "False positive validation process, validation records",
            "category": "DAST Remediation"
        },
        {
            "id": "3.9",
            "requirement": "API security testing tools are deployed for API testing",
            "evidence": "API testing tool documentation (Postman, REST Assured, etc.)",
            "category": "API Testing"
        },
        {
            "id": "3.10",
            "requirement": "API tests include authentication and authorisation testing",
            "evidence": "API test scenarios, auth bypass test cases",
            "category": "API Testing"
        },
        {
            "id": "3.11",
            "requirement": "API tests include input validation and injection testing",
            "evidence": "API fuzzing results, injection test cases",
            "category": "API Testing"
        },
        {
            "id": "3.12",
            "requirement": "Container/image scanning tool is deployed",
            "evidence": "Tool documentation (Trivy, Clair, Anchore), scan integration",
            "category": "Container Security"
        },
        {
            "id": "3.13",
            "requirement": "Container images are scanned before deployment",
            "evidence": "Container scan results, vulnerability reports",
            "category": "Container Security"
        },
        {
            "id": "3.14",
            "requirement": "IaC (Infrastructure-as-Code) scanning tool is deployed",
            "evidence": "Tool documentation (Checkov, tfsec, Terrascan), scan config",
            "category": "IaC Security"
        },
        {
            "id": "3.15",
            "requirement": "IaC is scanned for misconfigurations before deployment",
            "evidence": "IaC scan results, misconfiguration findings",
            "category": "IaC Security"
        },
        {
            "id": "3.16",
            "requirement": "Security testing results are archived for audit purposes",
            "evidence": "Test result storage, historical scan archives",
            "category": "Testing Governance"
        },
        {
            "id": "3.17",
            "requirement": "Penetration testing is performed annually for critical applications",
            "evidence": "Pentest reports, remediation tracking, retest results",
            "category": "Penetration Testing"
        },
        {
            "id": "3.18",
            "requirement": "Bug bounty or responsible disclosure program is active",
            "evidence": "Bug bounty platform (HackerOne, Bugcrowd), submission records",
            "category": "External Testing"
        },
    ]


def get_domain4_requirements():
    """
    Domain 4: IDE Plugins & Linters
    
    Catching issues in the IDE is the earliest and cheapest place to fix them.
    Developer-facing tools need high usability or they'll be ignored.
    """
    return [
        {
            "id": "4.1",
            "requirement": "Security-focused IDE plugins are available for developers",
            "evidence": "Plugin catalog, installation guides, approved plugin list",
            "category": "IDE Plugins"
        },
        {
            "id": "4.2",
            "requirement": "SAST IDE plugins provide real-time security feedback",
            "evidence": "SonarLint, Snyk IDE plugin, or similar tool documentation",
            "category": "IDE Plugins"
        },
        {
            "id": "4.3",
            "requirement": "IDE plugins are integrated with organisational security policies",
            "evidence": "Plugin configuration synced with SAST/SCA rules",
            "category": "IDE Configuration"
        },
        {
            "id": "4.4",
            "requirement": "IDE plugin adoption rate is measured and tracked",
            "evidence": "Plugin telemetry, adoption metrics dashboard",
            "category": "IDE Adoption"
        },
        {
            "id": "4.5",
            "requirement": "Developers receive training on using IDE security plugins",
            "evidence": "Plugin training materials, usage documentation",
            "category": "IDE Training"
        },
        {
            "id": "4.6",
            "requirement": "Code linters are configured with security rules",
            "evidence": "Linter configuration files (eslint, pylint, rubocop), security rules enabled",
            "category": "Linting"
        },
        {
            "id": "4.7",
            "requirement": "Linters run automatically in developer IDEs",
            "evidence": "IDE linter integration, pre-commit hook configuration",
            "category": "Linting"
        },
        {
            "id": "4.8",
            "requirement": "Linting rules are enforced in CI/CD pipeline",
            "evidence": "Pipeline linting step, linting failure logs",
            "category": "Linting"
        },
        {
            "id": "4.9",
            "requirement": "Secret scanning pre-commit hooks are deployed",
            "evidence": "Git hooks configuration (git-secrets, detect-secrets), hook installation",
            "category": "Secret Prevention"
        },
        {
            "id": "4.10",
            "requirement": "Pre-commit hooks prevent commits containing secrets",
            "evidence": "Blocked commit logs, secret detection examples",
            "category": "Secret Prevention"
        },
        {
            "id": "4.11",
            "requirement": "Developers can bypass pre-commit hooks only with justification",
            "evidence": "Bypass process documentation, bypass audit logs",
            "category": "Secret Prevention"
        },
        {
            "id": "4.12",
            "requirement": "Code formatting tools enforce consistent style",
            "evidence": "Prettier, Black, gofmt configuration, formatting enforcement",
            "category": "Code Quality"
        },
        {
            "id": "4.13",
            "requirement": "Complexity metrics are tracked (cyclomatic complexity, etc.)",
            "evidence": "SonarQube complexity metrics, code quality dashboards",
            "category": "Code Quality"
        },
        {
            "id": "4.14",
            "requirement": "Code coverage metrics are tracked and enforced",
            "evidence": "Coverage reports, minimum coverage thresholds",
            "category": "Code Quality"
        },
        {
            "id": "4.15",
            "requirement": "Dependency update tools provide automated update PRs",
            "evidence": "Dependabot, Renovate Bot configuration, automated PRs",
            "category": "Dependency Management"
        },
        {
            "id": "4.16",
            "requirement": "IDE plugins provide inline documentation for secure coding",
            "evidence": "Context-aware security hints, inline documentation",
            "category": "IDE Plugins"
        },
        {
            "id": "4.17",
            "requirement": "Developer feedback on tool usability is collected and acted upon",
            "evidence": "Developer surveys, tool improvement backlog",
            "category": "Tool Governance"
        },
        {
            "id": "4.18",
            "requirement": "Tool performance impact on developer workflow is monitored",
            "evidence": "Performance metrics, developer satisfaction scores",
            "category": "Tool Governance"
        },
    ]


def get_domain5_requirements():
    """
    Domain 5: Tool Effectiveness & Metrics
    
    Tools are investments. We need to measure if they're working.
    As Feynman said: "If it disagrees with experiment, it's wrong."
    """
    return [
        {
            "id": "5.1",
            "requirement": "Security tool effectiveness is measured with KPIs",
            "evidence": "KPI dashboard, effectiveness metrics definition",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.2",
            "requirement": "Vulnerability detection rate is tracked per tool",
            "evidence": "Detection metrics by tool, true positive rates",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.3",
            "requirement": "False positive rate is measured and minimised",
            "evidence": "False positive metrics, tuning efforts documentation",
            "category": "Tool Tuning"
        },
        {
            "id": "5.4",
            "requirement": "Mean time to remediation (MTTR) is tracked for findings",
            "evidence": "MTTR metrics, remediation velocity dashboard",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.5",
            "requirement": "Tool coverage is measured (% of codebase scanned)",
            "evidence": "Coverage reports, unscanned code identification",
            "category": "Tool Coverage"
        },
        {
            "id": "5.6",
            "requirement": "Security tool findings trends are analysed over time",
            "evidence": "Trend analysis reports, historical metrics",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.7",
            "requirement": "Tool costs are tracked and justified by value delivered",
            "evidence": "Cost-benefit analysis, ROI documentation",
            "category": "Tool Governance"
        },
        {
            "id": "5.8",
            "requirement": "Tool performance (scan time, resource usage) is monitored",
            "evidence": "Performance monitoring, scan duration metrics",
            "category": "Tool Performance"
        },
        {
            "id": "5.9",
            "requirement": "Duplicate findings across tools are identified and deduplicated",
            "evidence": "Deduplication process, consolidated findings reports",
            "category": "Tool Integration"
        },
        {
            "id": "5.10",
            "requirement": "Security findings are correlated with production incidents",
            "evidence": "Incident analysis, finding-to-incident correlation reports",
            "category": "Metrics & Measurement"
        },
        {
            "id": "5.11",
            "requirement": "Tool findings are integrated with vulnerability management system",
            "evidence": "VM system integration, unified vulnerability tracking",
            "category": "Tool Integration"
        },
        {
            "id": "5.12",
            "requirement": "Security metrics are reported to management quarterly",
            "evidence": "Management reports, executive dashboards",
            "category": "Reporting"
        },
        {
            "id": "5.13",
            "requirement": "Tool configuration changes are tracked and audited",
            "evidence": "Configuration change logs, audit trails",
            "category": "Tool Governance"
        },
        {
            "id": "5.14",
            "requirement": "Tool licenses are managed and renewal dates tracked",
            "evidence": "License management system, renewal calendar",
            "category": "Tool Governance"
        },
        {
            "id": "5.15",
            "requirement": "New tools are evaluated against security requirements",
            "evidence": "Tool evaluation framework, POC results",
            "category": "Tool Selection"
        },
        {
            "id": "5.16",
            "requirement": "Tools are regularly updated to latest stable versions",
            "evidence": "Tool version inventory, update schedule",
            "category": "Tool Maintenance"
        },
        {
            "id": "5.17",
            "requirement": "Tool effectiveness is benchmarked against industry standards",
            "evidence": "Benchmark comparison reports, industry metrics",
            "category": "Tool Effectiveness"
        },
        {
            "id": "5.18",
            "requirement": "Continuous improvement process exists for tool optimization",
            "evidence": "Improvement backlog, optimization initiatives",
            "category": "Tool Governance"
        },
    ]

# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET CREATION
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each assessment domain: Coding Standards, SAST/SCA, DAST, IDE Plugins, Tool Metrics.', '2. Assess DEPLOYMENT (is it there?) and EFFECTIVENESS (does it work?).', '3. For tools: assess whether they are configured and used properly, not just deployed.', '4. Provide specific tool names and versions as evidence.', '5. Reference actual scan results, configuration files, and adoption metrics.', '6. Use Tool Effectiveness Metrics sheet to track measurable outcomes.', '7. Document gaps with specific remediation actions.', '8. Obtain approvals from both Security and Development leadership.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_domain_sheet(ws, domain_name, requirements, styles):
    """Create standardised assessment domain sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Standards & Tools Assessment: {domain_name}".upper()
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO 27001:2022 | Control A.8.28 | {domain_name}"
    ws["A2"].font = Font(italic=True, size=10, color="003366", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["ID", "Requirement", "Implementation Status", "Evidence Reference", "Comments"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        apply_style(cell, styles["column_header"])
    
    ws.row_dimensions[3].height = 30
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40

    current_category = None
    row = 4
    
    for req in requirements:
        if "category" in req and req["category"] != current_category:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"▶ {req['category']}"
            ws[f"A{row}"].font = Font(bold=True, size=11, color="4472C4")
            ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            current_category = req["category"]
        
        ws[f"A{row}"] = req["id"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        ws[f"B{row}"] = req["requirement"]
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        validations["implementation_status"].add(ws[f"C{row}"])
        
        ws[f"D{row}"] = req.get("evidence", "")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"D{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[f"D{row}"].font = Font(size=9, italic=True, color="666666")
        
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"E{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row += 1

    # Pad to 50 FFFFCC data rows for QA compliance (DS-005)
    min_data_rows = 50
    data_rows_written = len(requirements)
    if data_rows_written < min_data_rows:
        thin = Side(style="thin")
        for pad_row in range(row, row + (min_data_rows - data_rows_written)):
            for col in ["A", "B", "C", "D", "E"]:
                cell = ws[f"{col}{pad_row}"]
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws[f"C{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"C{pad_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            validations["implementation_status"].add(ws[f"C{pad_row}"])
            ws[f"D{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"D{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws[f"E{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"E{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Hide unused columns (F onwards) to prevent empty column display
    for col_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    return ws


def add_grey_info_rows(ws, items):
    """
    Add F2F2F2 grey informational rows after the last requirement row.

    These rows are visible reference information and are NOT counted by COUNTIF
    formulas (which target FFFFCC input cells only). They overwrite the first
    N FFFFCC padding rows written by create_domain_sheet().

    Strategy: find the last row where column B has content (last requirement),
    then the next row is the first FFFFCC padding row — overwrite with F2F2F2.
    """
    thin = Side(style="thin")
    # Find last row with content in column B (last requirement row)
    last_req_row = 2  # default to header row
    for row_idx in range(ws.max_row, 2, -1):
        cell_b = ws.cell(row=row_idx, column=2)
        if cell_b.value is not None and str(cell_b.value).strip() != "":
            last_req_row = row_idx
            break

    # Overwrite the first len(items) FFFFCC padding rows with F2F2F2
    for i, item_text in enumerate(items):
        r = last_req_row + 1 + i
        for col in range(1, 6):  # columns A-E
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.cell(row=r, column=2).value = item_text
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, color="003366", italic=True)
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        # Clear any FFFFCC fill or DV from col C (it was a padding input cell)
        ws.cell(row=r, column=3).fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )


def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard matching A.8.33 gold standard with TABLE 1, 2, 3."""
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (italic, 003366 font, no fill)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 Annex A.8.28 — Secure Coding | Standards & Tools Assessment"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4: TABLE 1 banner (003366, merged A4:G4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=4, column=col).border = border

    # Row 5: Column headers (D9D9D9 grey, black text)
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows - assessment areas
    area_configs = [
        ("Coding Standards Adoption", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("SAST SCA Tools", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("DAST Security Testing Tools", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("IDE Plugins Linters", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Tool Effectiveness Metrics", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
    ]

    for i, (area_name, status_col, status_values) in enumerate(area_configs):
        row = 6 + i
        ws.cell(row=row, column=1, value=area_name).border = border
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTA('{area_name}'!{status_col}4:{status_col}100)"
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal="center")
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[0]}")'
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal="center")
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[1]}")'
        cell_d.border = border
        cell_d.alignment = Alignment(horizontal="center")
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[2]}")'
        cell_e.border = border
        cell_e.alignment = Alignment(horizontal="center")
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"N/A")'
        cell_f.border = border
        cell_f.alignment = Alignment(horizontal="center")
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f'=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))'
        cell_g.number_format = "0.0%"
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center")

    # TOTAL row (D9D9D9 fill, data rows 6–N)
    total_row = 6 + len(area_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1).border = border
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    cell = ws.cell(row=total_row, column=7)
    cell.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    cell.number_format = "0.0%"
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    # TABLE 2: Key Metrics
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{metrics_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=metrics_start, column=col).border = border

    # TABLE 2 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        hcell = ws.cell(row=metrics_start + 1, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center")

    metrics = [
        # Coding Standards
        ("Total Coding Standards", "=COUNTA('Coding Standards Adoption'!B5:B100)"),
        ("Standards Fully Adopted", '=COUNTIF(\'Coding Standards Adoption\'!C5:C100,"✅ Implemented")'),
        ("Standards Adoption Rate", '=COUNTIF(\'Coding Standards Adoption\'!C5:C100,"✅ Implemented")/COUNTA(\'Coding Standards Adoption\'!B5:B100)'),
        # SAST/SCA Tools
        ("Total SAST/SCA Tools", "=COUNTA('SAST SCA Tools'!B5:B100)"),
        ("Tools Fully Integrated", '=COUNTIF(\'SAST SCA Tools\'!C5:C100,"✅ Implemented")'),
        ("SAST/SCA Coverage", '=COUNTIF(\'SAST SCA Tools\'!C5:C100,"✅ Implemented")/COUNTA(\'SAST SCA Tools\'!B5:B100)'),
        # DAST Tools
        ("Total DAST Tools", "=COUNTA('DAST Security Testing Tools'!B5:B100)"),
        ("DAST Tools Active", '=COUNTIF(\'DAST Security Testing Tools\'!C5:C100,"✅ Implemented")'),
        ("DAST Testing Coverage", '=COUNTIF(\'DAST Security Testing Tools\'!C5:C100,"✅ Implemented")/COUNTA(\'DAST Security Testing Tools\'!B5:B100)'),
        # IDE Plugins
        ("Total IDE Plugins", "=COUNTA('IDE Plugins Linters'!B5:B100)"),
        ("Plugins Deployed", '=COUNTIF(\'IDE Plugins Linters\'!C5:C100,"✅ Implemented")'),
        ("Plugin Adoption Rate", '=COUNTIF(\'IDE Plugins Linters\'!C5:C100,"✅ Implemented")/COUNTA(\'IDE Plugins Linters\'!B5:B100)'),
        # Effectiveness
        ("Total Effectiveness Metrics", "=COUNTA('Tool Effectiveness Metrics'!B5:B100)"),
        ("Metrics Tracked", '=COUNTIF(\'Tool Effectiveness Metrics\'!C5:C100,"✅ Implemented")'),
        ("Overall Tool Effectiveness", '=COUNTIF(\'Tool Effectiveness Metrics\'!C5:C100,"✅ Implemented")/COUNTA(\'Tool Effectiveness Metrics\'!B5:B100)'),
    ]

    r = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=r, column=1, value=metric).border = border
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10)
        cell_val = ws.cell(row=r, column=2, value=formula)
        cell_val.border = border
        cell_val.font = Font(name="Calibri", size=10, color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # 2 buffer rows after TABLE 2 metrics
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # TABLE 3: Critical Findings
    crit_start = r + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    # TABLE 3 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Coding Standards Adoption", "No coding standards enforced", '=COUNTIF(\'Coding Standards Adoption\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("SAST SCA Tools", "Missing SAST/SCA in CI/CD pipeline", '=COUNTIF(\'SAST SCA Tools\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("DAST Security Testing Tools", "Zero DAST coverage for APIs", '=COUNTIF(\'DAST Security Testing Tools\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("IDE Plugins Linters", "No security plugins in developer IDEs", '=COUNTIF(\'IDE Plugins Linters\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Tool Effectiveness Metrics", "No metrics tracking tool effectiveness", '=COUNTIF(\'Tool Effectiveness Metrics\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Coding Standards Adoption", "Inconsistent standards adoption", '=COUNTIF(\'Coding Standards Adoption\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("SAST SCA Tools", "Incomplete SAST/SCA integration", '=COUNTIF(\'SAST SCA Tools\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("DAST Security Testing Tools", "Partial DAST implementation", '=COUNTIF(\'DAST Security Testing Tools\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("IDE Plugins Linters", "Low plugin adoption rates", '=COUNTIF(\'IDE Plugins Linters\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Tool Effectiveness Metrics", "Incomplete metrics collection", '=COUNTIF(\'Tool Effectiveness Metrics\'!C5:C100,"⚠️ Partially Implemented")', "Medium", "Plan"),
    ]

    r3 = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        ws.cell(row=r3, column=1, value=cat)
        ws.cell(row=r3, column=2, value=finding)
        count_cell = ws.cell(row=r3, column=3, value=formula)
        count_cell.alignment = Alignment(horizontal="center")
        count_cell.font = Font(name="Calibri", size=10, color="000000")
        ws.cell(row=r3, column=4, value=severity)
        ws.cell(row=r3, column=5, value=action)
        r3 += 1

    # 2 buffer rows after TABLE 3
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        r3 += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"

    return ws


def create_evidence_register(ws, styles):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        ("Evidence ID", 15),
        ("Assessment Area", 25),
        ("Evidence Type", 22),
        ("Description", 40),
        ("Location/Path", 45),
        ("Date Collected", 16),
        ("Collected By", 20),
        ("Verification Status", 22),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border_thin
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Data rows EV-001 to EV-100
    dv_etype = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=False
    )
    dv_vstatus = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False
    )
    ws.add_data_validation(dv_etype)
    ws.add_data_validation(dv_vstatus)

    # MAX-001/MAX-002 fix: 1 grey sample + 100 FFFFCC empty rows
    # Sample row (row 5) with F2F2F2 grey background (like A.8.33 reference)
    sample_data = [
        "EV-001",
        "Test Data Inventory",
        "Configuration file",
        "Data masking configuration for UAT environment",
        "\\fileserver\\isms\\evidence\\test-data\\uat-masking-config.json",
        "2024-03-01",
        "Sarah Mitchell",
        "Verified"
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=5, column=col_num, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = border_thin

    # 100 empty rows (rows 6-105) - FFFFCC yellow
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
        dv_etype.add(ws.cell(row=r, column=3))
        dv_vstatus.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"
    return ws


def create_gap_analysis_sheet(ws, styles):
    """Create Gap Analysis sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:J1")
    ws["A1"] = "GAP ANALYSIS & REMEDIATION PLAN"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all 'Not Implemented' and 'Partially Implemented' requirements. Assign owners, set target dates, and track remediation progress."
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Gap ID", "Domain", "Requirement ID", "Requirement Description", "Current State", "Target State", "Priority", "Owner", "Target Date", "Status"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    widths = {"A": 10, "B": 22, "C": 12, "D": 35, "E": 25, "F": 25, "G": 12, "H": 20, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    validations["priority"].add("G4:G200")
    validations["gap_status"].add("J4:J200")

    # MAX-003 fix: 1 sample row (F2F2F2 grey) + 50 empty rows (FFFFCC yellow) per Option B standard
    thin = Side(style="thin")

    # Sample row (row 4) with realistic example data
    sample_data = [
        "GAP-001",
        "SAST & SCA Tools",
        "2.2",
        "SonarQube integrated into CI/CD pipeline",
        "SAST runs manually outside pipeline",
        "Automated SAST scans with quality gates in Jenkins",
        "High",
        "DevSecOps Lead",
        "28.02.2025",
        "In Progress"
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col_num, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 50 empty rows (rows 5-54) - FFFFCC yellow
    for gap_row in range(5, 55):
        for col in range(1, 11):
            cell = ws.cell(row=gap_row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Hide unused columns (K onwards) to prevent empty column display in Gap Analysis
    for col_letter in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    return ws


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=1, column=col).border = border_thin
    ws.row_dimensions[1].height = 35

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=2, column=col).border = border_thin

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=row, column=col).border = border_thin

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.28.2 - Standards & Tools Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        # Apply borders to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = border_thin
            if value == "":
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Status dropdown on Assessment Status
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Apply borders to banner (A:E)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_thin
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            # Apply borders and fill to all cells in merged range (B:E)
            for col in range(2, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=row, column=col).border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all cells in merged range (B:E)
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border_thin

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to banner (A:E)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border_thin

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

    return ws



def main():
    """Generate Standards & Tools Assessment workbook."""
    logger.info("=" * 80)
    logger.info(" " * 15 + "ISMS Control 8.28.2 - Standards & Tools Assessment Generator")
    logger.info("=" * 80)
    logger.info("")
    
    try:
        logger.info("[N] [1/10] Creating workbook structure...")
        wb = create_workbook()
        styles = _STYLES
        logger.info("     ✓ Workbook initialized with 10 sheets")

        logger.info("[F] [2/10] Creating Instructions sheet...")
        ws_instructions = wb["Instructions & Legend"]
        ws_instructions.sheet_view.showGridLines = False
        create_instructions_sheet(ws_instructions)
        logger.info("     ✓ Instructions complete (focus: deployment AND effectiveness)")

        logger.info("[CHART] [3/10] Creating Domain 1: Coding Standards Adoption...")
        ws_domain1 = wb["Coding Standards Adoption"]
        ws_domain1.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain1, "Coding Standards Adoption", get_domain1_requirements(), styles)
        # ISO 27002:2022 Phase 2 content: add banned/unsafe construct reference rows (F2F2F2)
        add_grey_info_rows(ws_domain1, [
            "Banned/unsafe functions list maintained — language-specific unsafe APIs prohibited "
            "(e.g. strcpy/strcat in C, eval() in Python/JS, pickle.loads() in Python, exec() in PHP)",
            "Memory safety requirements defined for C/C++ code — buffer overflow prevention "
            "(bounds checking, safe libraries), use-after-free detection",
            "Integer overflow/underflow protections required for arithmetic on untrusted input",
        ])
        logger.info("     \u2713 18 requirements + 3 reference rows (Standards, Training, Enforcement)")

        logger.info("[CHART] [4/10] Creating Domain 2: SAST & SCA Tools...")
        ws_domain2 = wb["SAST SCA Tools"]
        ws_domain2.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain2, "SAST & SCA Tools", get_domain2_requirements(), styles)
        logger.info("     ✓ 18 requirements (Static Analysis, Dependency Scanning)")

        logger.info("[CHART] [5/10] Creating Domain 3: DAST & Security Testing Tools...")
        ws_domain3 = wb["DAST Security Testing Tools"]
        ws_domain3.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain3, "DAST & Security Testing Tools", get_domain3_requirements(), styles)
        logger.info("     ✓ 18 requirements (Dynamic Testing, API Testing, Container/IaC)")

        logger.info("[CHART] [6/10] Creating Domain 4: IDE Plugins & Linters...")
        ws_domain4 = wb["IDE Plugins Linters"]
        ws_domain4.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain4, "IDE Plugins & Linters", get_domain4_requirements(), styles)
        logger.info("     ✓ 18 requirements (IDE Plugins, Linters, Pre-commit Hooks)")

        logger.info("[CHART] [7/10] Creating Domain 5: Tool Effectiveness & Metrics...")
        ws_domain5 = wb["Tool Effectiveness Metrics"]
        ws_domain5.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain5, "Tool Effectiveness & Metrics", get_domain5_requirements(), styles)
        logger.info("     ✓ 18 requirements (KPIs, Coverage, Remediation Velocity)")

        logger.info("[TREND] [8/10] Creating Summary Dashboard...")
        ws_summary = wb["Summary Dashboard"]
        ws_summary.sheet_view.showGridLines = False
        create_summary_dashboard_sheet(ws_summary, styles)
        logger.info("     ✓ Executive summary with tool effectiveness metrics")

        logger.info("[9/10] Creating Evidence Register...")
        ws_evidence = wb["Evidence Register"]
        ws_evidence.sheet_view.showGridLines = False
        create_evidence_register(ws_evidence, styles)
        logger.info("     ✓ Evidence tracking")

        logger.info("\u26A0\uFE0F  [10/10] Creating Gap Analysis & Approval sheets...")
        ws_gap = wb["Gap Analysis"]
        ws_gap.sheet_view.showGridLines = False
        create_gap_analysis_sheet(ws_gap, styles)
        
        ws_approval = wb["Approval Sign-Off"]
        ws_approval.sheet_view.showGridLines = False
        create_approval_sheet(ws_approval, styles)
        logger.info("     ✓ Gap tracking and approval workflow")

        logger.info("")
        logger.info("[S] Saving workbook...")
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.8.28.2_Standards_Tools_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info("")
        logger.info("=" * 80)
        logger.info("\u2705 SUCCESS: Standards & Tools Assessment workbook generated!")
        logger.info("=" * 80)
        logger.info("")
        logger.info(f"[D] File: {filename}")
        logger.info(f"[CHART] Total Requirements: 90 (18 per domain × 5 domains)")
        logger.info("")
        logger.info("[PIN] Assessment Domains:")
        logger.info("   1. Coding Standards Adoption          (18 requirements)")
        logger.info("   2. SAST & SCA Tools                   (18 requirements)")
        logger.info("   3. DAST & Security Testing Tools      (18 requirements)")
        logger.info("   4. IDE Plugins & Linters              (18 requirements)")
        logger.info("   5. Tool Effectiveness & Metrics       (18 requirements)")
        logger.info("")
        logger.info("[TIP] Key Focus: Tools are investments. Measure if they're working.")
        logger.info("   Don't just check if tools exist - verify they're effective!")
        logger.info("")
        logger.info("=" * 80)
        
        return 0

    except Exception as e:
        logger.info("")
        logger.info("=" * 80)
        logger.error("\u274C ERROR: Failed to generate workbook")
        logger.info("=" * 80)
        logger.error(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
