#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.28.3 - Code Review & Testing Assessment
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Assessment Domain 3 of 4: Code Review and Security Testing Processes

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific code review practices, testing methodologies,
and quality assurance processes.

Key customisation areas:
1. Code review workflows and tools (match your review process and platforms)
2. Security testing types and coverage (adapt to your testing strategy)
3. Testing tools and frameworks (specific to your technology stack)
4. Review criteria and quality gates (based on your quality standards)
5. Compliance thresholds and scoring (aligned with your risk tolerance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMISE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
code review effectiveness and security testing coverage across all development
projects and repositories.

**Purpose:**
Enables systematic assessment of code review and security testing processes
against ISO 27001:2022 Control A.8.28 requirements, supporting evidence-based
validation of quality assurance and vulnerability detection practices.

**Assessment Scope:**
- Peer code review processes and coverage metrics
- Security-focused code review checklists and criteria
- Static Application Security Testing (SAST) integration and coverage
- Dynamic Application Security Testing (DAST) implementation
- Software Composition Analysis (SCA) for dependencies
- Interactive Application Security Testing (IAST) where applicable
- Penetration testing frequency and scope
- Security regression testing and test case management
- Code coverage metrics and security test coverage
- Vulnerability management and remediation tracking
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and testing standard references
2. Peer Review - Code review process effectiveness and coverage assessment
3. Security Review - Security-focused review criteria and execution
4. SAST Coverage - Static analysis implementation and finding resolution
5. DAST Testing - Dynamic testing coverage and vulnerability detection
6. SCA Analysis - Dependency scanning and vulnerable component management
7. Penetration Testing - Pen test frequency, scope, and finding remediation
8. Test Coverage - Code coverage and security test case adequacy
9. Vulnerability Mgmt - Defect tracking, prioritization, and remediation SLAs
10. Gap Analysis - Inadequate review/testing coverage and remediation
11. Evidence Register - Audit evidence tracking and documentation
12. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with testing methodology dropdown lists
- Conditional formatting for coverage thresholds and compliance status
- Automated gap identification for untested code or missing test types
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with test management and defect tracking systems

**Integration:**
This assessment feeds into the A.8.28.5 Compliance Dashboard, which
consolidates data from all four assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a828_3_code_review_testing.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a828_3_code_review_testing.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a828_3_code_review_testing.py --date 20250124

Output:
    File: ISMS_A_8_28_3_Code_Review_Testing_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise testing standards to match your QA methodology
    2. Inventory all code repositories, projects, and applications
    3. Complete assessments for each project or development team
    4. Validate SAST/DAST/SCA tool coverage and scan frequency
    5. Review penetration testing scope and remediation timelines
    6. Conduct gap analysis for insufficient testing or review coverage
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (scan reports, review records, pen test results)
    9. Obtain stakeholder approvals (QA Lead, AppSec, Dev Manager)
    10. Feed results into A.8.28.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    3 of 4 (Code Review & Testing)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Implementation Guide
    - ISMS-IMP-A.8.28.1: SDLC Integration Assessment (Domain 1)
    - ISMS-IMP-A.8.28.2: Standards & Tools Assessment (Domain 2)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Assessment (Domain 4)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a828_1_sdlc_assessment.py
    - generate_a828_2_standards_tools.py
    - generate_a828_4_third_party_oss.py
    - generate_a828_5_compliance_dashboard.py
    - normalize_assessment_files_a828.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.28.3 specification
    - Supports comprehensive code review and security testing evaluation
    - Integrated with A.8.28.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Testing Strategy Balance:**
Effective application security requires layered testing approaches. Don't
rely solely on automated tools (SAST/DAST/SCA) - they miss context-specific
vulnerabilities. Don't rely solely on manual testing - it doesn't scale.
Balance is key.

**Code Review Effectiveness:**
Code review participation rates don't equal effectiveness. A rushed review
that approves everything is worse than no review. Assess quality of reviews:
- Are security issues actually caught in review?
- How long do reviews take on average?
- What percentage result in meaningful feedback vs. rubber-stamping?

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of testing coverage, review records, and
vulnerability remediation timelines.

**Data Protection:**
Assessment workbooks contain sensitive security information including:
- Vulnerability details and security defect patterns
- Penetration test findings and attack scenarios
- Code coverage gaps and untested functionality
- Security testing tool configurations and results

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Review testing coverage and vulnerability remediation rates
- Semi-annually: Update testing methodologies for new threat patterns
- Annually: Complete reassessment of all projects and applications
- Ad-hoc: When new testing tools adopted or testing requirements change

**Quality Assurance:**
Have application security SMEs, QA leads, and security testing specialists
validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
Ensure testing standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 testing requirements (Requirement 6.3, 11.3)
- Healthcare: HIPAA security testing and vulnerability management
- Finance: Regional banking penetration testing requirements
- Government: Jurisdiction-specific security testing mandates

Customise assessment criteria to include regulatory-specific requirements.

**Integration with Testing Tools:**
Where possible, integrate assessment data with:
- Code review platforms (GitHub, GitLab, Bitbucket, Gerrit, etc.)
- Test management systems (Jira, Azure DevOps, TestRail, etc.)
- SAST/DAST/SCA tool outputs (detailed scan results and trends)
- Defect tracking systems (vulnerability lifecycle and remediation metrics)
- CI/CD pipelines (automated test execution and coverage reports)

Automated tool integration reduces manual effort and improves accuracy.

**Don't Fool Yourself - Feynman Principle:**
High test coverage percentages don't automatically mean good security testing.
100% code coverage with zero security test cases is worthless. Similarly,
running SAST scans doesn't help if findings aren't reviewed and addressed.

Assess actual security defect detection and remediation:
- How many security vulnerabilities were found in the last 6 months?
- What percentage were found by automated vs. manual testing?
- How quickly are high/critical findings remediated?
- Can you show examples of security issues caught before production?

If you can't demonstrate defects found and fixed, your testing isn't working.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.28.3"
WORKBOOK_NAME = "Code Review and Security Testing Processes"
CONTROL_ID = "A.8.28"
CONTROL_NAME = "Secure Coding"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


import sys

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.28.3 specification
    sheets = [
        "Instructions & Legend",
        "Code Review Process",
        "Security Champion Review",
        "Unit Integration Testing",
        "API Application Testing",
        "External Testing Validation",
        "Evidence Register",
        "Gap Analysis",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Code review and testing are the last lines of defense before production.
    This workbook helps assess whether those defenses are actually working.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_implemented": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notimplemented": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        },
        "priority_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "priority_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "priority_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "priority_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell with fresh objects."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    
    Process without effectiveness is theater. These validations help
    assess whether reviews and tests actually catch vulnerabilities.
    """
    validations = {
        # Implementation status
        'implementation_status': DataValidation(
            type="list",
            formula1='"✅ Implemented,⚠️ Partially Implemented,❌ Not Implemented,N/A"',
            allow_blank=False
        ),
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Unknown"',
            allow_blank=False
        ),
        
        # Review and testing effectiveness
        'effectiveness_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Fair,Poor,Not Assessed"',
            allow_blank=False
        ),
        'coverage_level': DataValidation(
            type="list",
            formula1='"Complete (>95%),High (80-95%),Medium (60-79%),Low (<60%),Unknown"',
            allow_blank=False
        ),
        'frequency': DataValidation(
            type="list",
            formula1='"Every Change,Daily,Weekly,Monthly,Quarterly,On-Demand,Never"',
            allow_blank=False
        ),
        
        # Priority and risk
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Negligible"',
            allow_blank=False
        ),
        
        # Gap status
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Closed,Deferred"',
            allow_blank=False
        ),
        
        # Evidence types
        'evidence_type': DataValidation(
            type="list",
            formula1='"Review Log,Test Result,Pentest Report,Bug Bounty Submission,Training Record,Policy Document,Tool Output,Screenshot,Other"',
            allow_blank=False
        ),
        
        # Review types
        'review_type': DataValidation(
            type="list",
            formula1='"Peer Review,Security Champion Review,Architecture Review,Design Review,Post-Incident Review"',
            allow_blank=False
        ),
        
        # Test types
        'test_type': DataValidation(
            type="list",
            formula1='"Unit Test,Integration Test,API Test,Security Test,Penetration Test,Regression Test,Smoke Test"',
            allow_blank=False
        ),
        
        # Test status
        'test_status': DataValidation(
            type="list",
            formula1='"Passing,Failing,Flaky,Disabled,Not Run"',
            allow_blank=False
        ),
        
        # Automation level
        'automation_level': DataValidation(
            type="list",
            formula1='"Fully Automated,Partially Automated,Manual,Not Applicable"',
            allow_blank=False
        ),
        
        # Finding severity
        'severity': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Informational"',
            allow_blank=False
        ),
        
        # Remediation status
        'remediation_status': DataValidation(
            type="list",
            formula1='"Fixed,In Progress,Accepted Risk,False Positive,Deferred"',
            allow_blank=False
        ),
        
        # Compliance level
        'compliance_level': DataValidation(
            type="list",
            formula1='"Fully Compliant,Mostly Compliant,Partially Compliant,Non-Compliant,N/A"',
            allow_blank=False
        ),
        
        # Training status
        'training_status': DataValidation(
            type="list",
            formula1='"Completed,In Progress,Not Started,N/A"',
            allow_blank=False
        ),
        
        # Approval decision
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
            allow_blank=False
        ),
    }

    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: ASSESSMENT DOMAIN DATA DEFINITIONS
# ============================================================================

def get_domain1_requirements():
    """
    Domain 1: Code Review Process & Governance
    
    Code review is only effective if it's mandatory, done by qualified people,
    and actually catches security issues. "LGTM" is not a security review.
    """
    return [
        {
            "id": "1.1",
            "requirement": "Peer review is mandatory for all production code changes",
            "evidence": "Repository settings showing branch protection, review approval requirements",
            "category": "Review Policy"
        },
        {
            "id": "1.2",
            "requirement": "Code review policy is documented and accessible to all developers",
            "evidence": "Code review policy document, developer handbook, internal wiki",
            "category": "Review Policy"
        },
        {
            "id": "1.3",
            "requirement": "Minimum reviewer qualifications are defined and enforced",
            "evidence": "Reviewer qualification criteria, approved reviewer list",
            "category": "Review Policy"
        },
        {
            "id": "1.4",
            "requirement": "All reviewers are trained on secure code review practices",
            "evidence": "Training completion records, secure code review training materials",
            "category": "Reviewer Training"
        },
        {
            "id": "1.5",
            "requirement": "Security-focused checklist is used in all code reviews",
            "evidence": "Code review checklist template, checklist usage in PR comments",
            "category": "Review Process"
        },
        {
            "id": "1.6",
            "requirement": "Code review tools are integrated into development workflow",
            "evidence": "GitHub/GitLab/Bitbucket PR process, review tool configuration",
            "category": "Review Tools"
        },
        {
            "id": "1.7",
            "requirement": "Review approval is required before code can be merged to production",
            "evidence": "Branch protection rules, merge approval logs",
            "category": "Review Enforcement"
        },
        {
            "id": "1.8",
            "requirement": "Code review coverage is measured and tracked",
            "evidence": "Review coverage metrics dashboard, coverage reports",
            "category": "Review Metrics"
        },
        {
            "id": "1.9",
            "requirement": "Review comments reference secure coding standards and best practices",
            "evidence": "Sample review comments with standards references",
            "category": "Review Quality"
        },
        {
            "id": "1.10",
            "requirement": "High-risk code changes require additional security review scrutiny",
            "evidence": "Risk-based review policy, high-risk change review records",
            "category": "Review Process"
        },
        {
            "id": "1.11",
            "requirement": "Security findings from code review are tracked to resolution",
            "evidence": "Finding tracking system (JIRA, GitHub Issues), resolution logs",
            "category": "Review Follow-up"
        },
        {
            "id": "1.12",
            "requirement": "Code review metrics are reported to management quarterly",
            "evidence": "Quarterly review metrics reports, management dashboards",
            "category": "Review Metrics"
        },
        {
            "id": "1.13",
            "requirement": "Review bypass requires documented approval and justification",
            "evidence": "Exception request process, bypass approval records",
            "category": "Review Enforcement"
        },
        {
            "id": "1.14",
            "requirement": "Automated review assistance tools are deployed (PR bots, automated checks)",
            "evidence": "PR bot integration (Danger, CodeOwners), automated check configuration",
            "category": "Review Tools"
        },
        {
            "id": "1.15",
            "requirement": "Code review quality is audited periodically",
            "evidence": "Review quality audit results, sample review assessments",
            "category": "Review Quality"
        },
        {
            "id": "1.16",
            "requirement": "Self-review is not permitted for production code",
            "evidence": "Review policy prohibiting self-approval, enforcement evidence",
            "category": "Review Policy"
        },
        {
            "id": "1.17",
            "requirement": "Code review turnaround time is measured and within SLA",
            "evidence": "Review SLA metrics, turnaround time tracking",
            "category": "Review Metrics"
        },
        {
            "id": "1.18",
            "requirement": "Code review process is continuously improved based on lessons learned",
            "evidence": "Process improvement backlog, retrospective notes",
            "category": "Continuous Improvement"
        },
    ]


def get_domain2_requirements():
    """
    Domain 2: Security Champion Review & Architecture Review
    
    Security Champions are developers who become security force multipliers.
    They provide deeper security expertise where needed.
    """
    return [
        {
            "id": "2.1",
            "requirement": "Security Champion program is formally established",
            "evidence": "Program charter, Security Champion roster, program documentation",
            "category": "Champion Program"
        },
        {
            "id": "2.2",
            "requirement": "Security Champions are assigned to all development teams",
            "evidence": "Team assignments, Security Champion coverage matrix",
            "category": "Champion Program"
        },
        {
            "id": "2.3",
            "requirement": "Security Champions are trained on secure development and threat modeling",
            "evidence": "Training certifications, secure development course completion records",
            "category": "Champion Training"
        },
        {
            "id": "2.4",
            "requirement": "Criteria for Security Champion review are clearly defined",
            "evidence": "Review trigger criteria documentation (e.g., auth changes, crypto, PII)",
            "category": "Champion Process"
        },
        {
            "id": "2.5",
            "requirement": "High-risk code changes are automatically escalated to Security Champions",
            "evidence": "Escalation logs, Champion review records, automated escalation rules",
            "category": "Champion Process"
        },
        {
            "id": "2.6",
            "requirement": "Architecture review is required for all new systems and major changes",
            "evidence": "Architecture review policy, architecture review board meeting notes",
            "category": "Architecture Review"
        },
        {
            "id": "2.7",
            "requirement": "Security design review is performed before implementation begins",
            "evidence": "Design review meeting notes, security design sign-off records",
            "category": "Architecture Review"
        },
        {
            "id": "2.8",
            "requirement": "Threat modeling is performed for security-sensitive features",
            "evidence": "Threat model documents, STRIDE analysis, attack trees",
            "category": "Threat Modeling"
        },
        {
            "id": "2.9",
            "requirement": "Security Champion availability and response time are tracked and meet SLA",
            "evidence": "Response time SLA metrics, Champion availability data",
            "category": "Champion Performance"
        },
        {
            "id": "2.10",
            "requirement": "Security Champion recommendations are tracked to implementation",
            "evidence": "Recommendation tracking system, implementation verification",
            "category": "Champion Follow-up"
        },
        {
            "id": "2.11",
            "requirement": "Clear escalation path exists from Champions to AppSec team",
            "evidence": "Escalation procedures, escalation records, contact information",
            "category": "Champion Process"
        },
        {
            "id": "2.12",
            "requirement": "Security Champions have authority to block releases for security issues",
            "evidence": "Authority documentation, examples of Champion-blocked releases",
            "category": "Champion Authority"
        },
        {
            "id": "2.13",
            "requirement": "Security Champion program effectiveness is measured with KPIs",
            "evidence": "Program metrics dashboard, effectiveness assessment reports",
            "category": "Champion Metrics"
        },
        {
            "id": "2.14",
            "requirement": "Regular Security Champion sync meetings are held",
            "evidence": "Meeting schedules, attendance records, meeting notes",
            "category": "Champion Program"
        },
        {
            "id": "2.15",
            "requirement": "Security Champion knowledge base and playbooks are maintained",
            "evidence": "Internal wiki, Security Champion playbooks, decision trees",
            "category": "Champion Resources"
        },
        {
            "id": "2.16",
            "requirement": "New Security Champions are onboarded with systematic training plan",
            "evidence": "Champion onboarding checklist, training plan documentation",
            "category": "Champion Training"
        },
        {
            "id": "2.17",
            "requirement": "Security Champion contributions are recognised and rewarded",
            "evidence": "Recognition program documentation, awards, acknowledgments",
            "category": "Champion Program"
        },
        {
            "id": "2.18",
            "requirement": "Security Champions receive regular updates on emerging threats",
            "evidence": "Threat intelligence briefings, security newsletters, training sessions",
            "category": "Champion Training"
        },
    ]


def get_domain3_requirements():
    """
    Domain 3: Security Testing - Unit & Integration
    
    Security tests at the unit and integration level catch issues early.
    But only if they actually test security-relevant behavior.
    """
    return [
        {
            "id": "3.1",
            "requirement": "Security test cases exist for all critical application functionality",
            "evidence": "Security test suite, test case repository, test documentation",
            "category": "Test Coverage"
        },
        {
            "id": "3.2",
            "requirement": "Input validation is tested at the unit test level",
            "evidence": "Input validation unit tests, boundary testing, malicious input tests",
            "category": "Unit Testing"
        },
        {
            "id": "3.3",
            "requirement": "Authentication logic is thoroughly tested with unit tests",
            "evidence": "Authentication unit test suites, auth failure test cases",
            "category": "Unit Testing"
        },
        {
            "id": "3.4",
            "requirement": "Authorisation checks are tested at unit and integration levels",
            "evidence": "Authorisation test cases, access control test matrices",
            "category": "Unit Testing"
        },
        {
            "id": "3.5",
            "requirement": "Session management security is tested",
            "evidence": "Session security test cases, timeout testing, session fixation tests",
            "category": "Integration Testing"
        },
        {
            "id": "3.6",
            "requirement": "Cryptographic functions are tested for correct implementation",
            "evidence": "Cryptography unit tests, key generation tests, encryption/decryption tests",
            "category": "Unit Testing"
        },
        {
            "id": "3.7",
            "requirement": "Error handling is tested to prevent information disclosure",
            "evidence": "Error handling test scenarios, exception testing, stack trace prevention tests",
            "category": "Unit Testing"
        },
        {
            "id": "3.8",
            "requirement": "Integration tests validate end-to-end authentication/authorisation workflows",
            "evidence": "Integration test suites, workflow test scenarios",
            "category": "Integration Testing"
        },
        {
            "id": "3.9",
            "requirement": "Security test coverage is measured and tracked",
            "evidence": "Test coverage metrics for security-critical code, coverage reports",
            "category": "Test Metrics"
        },
        {
            "id": "3.10",
            "requirement": "Failed security tests block CI/CD pipeline deployment",
            "evidence": "Pipeline configuration, test failure blocking evidence, failed build logs",
            "category": "Test Enforcement"
        },
        {
            "id": "3.11",
            "requirement": "Security tests run automatically on every code commit",
            "evidence": "CI/CD configuration, test execution logs, automated test runs",
            "category": "Test Automation"
        },
        {
            "id": "3.12",
            "requirement": "Security test maintenance process exists and tests are kept current",
            "evidence": "Test maintenance procedures, test update logs, deprecation handling",
            "category": "Test Maintenance"
        },
        {
            "id": "3.13",
            "requirement": "Security test cases are peer-reviewed like production code",
            "evidence": "Test code review records, test PR approvals",
            "category": "Test Quality"
        },
        {
            "id": "3.14",
            "requirement": "Negative test cases are included for security-critical functions",
            "evidence": "Negative testing examples, failure mode testing",
            "category": "Unit Testing"
        },
        {
            "id": "3.15",
            "requirement": "Boundary condition testing is performed for input validation",
            "evidence": "Boundary test cases, edge case testing, overflow testing",
            "category": "Unit Testing"
        },
        {
            "id": "3.16",
            "requirement": "Race condition testing is performed where applicable",
            "evidence": "Concurrency test cases, threading tests, TOCTOU testing",
            "category": "Integration Testing"
        },
        {
            "id": "3.17",
            "requirement": "Security test results are archived for audit and trend analysis",
            "evidence": "Test result storage, historical test data, trend reports",
            "category": "Test Metrics"
        },
        {
            "id": "3.18",
            "requirement": "Test environment security is validated (no production data)",
            "evidence": "Test environment security assessment, data handling procedures",
            "category": "Test Environment"
        },
    ]


def get_domain4_requirements():
    """
    Domain 4: API & Application Security Testing
    
    APIs are attack surfaces. They need thorough security testing including
    authentication, authorisation, input validation, and abuse scenarios.
    """
    return [
        {
            "id": "4.1",
            "requirement": "All APIs (REST, GraphQL, gRPC) have dedicated security test suites",
            "evidence": "API security test documentation, Postman/REST Assured collections",
            "category": "API Testing"
        },
        {
            "id": "4.2",
            "requirement": "API authentication mechanisms are thoroughly tested",
            "evidence": "Auth testing scenarios (OAuth, JWT, API keys), token validation tests",
            "category": "API Testing"
        },
        {
            "id": "4.3",
            "requirement": "Authorisation is tested for all API endpoints and operations",
            "evidence": "Endpoint authorisation test matrix, RBAC testing, privilege escalation tests",
            "category": "API Testing"
        },
        {
            "id": "4.4",
            "requirement": "Input validation is tested for all API parameters",
            "evidence": "Input validation test cases, parameter fuzzing, type confusion tests",
            "category": "API Testing"
        },
        {
            "id": "4.5",
            "requirement": "SQL injection testing is performed on all database-backed APIs",
            "evidence": "SQL injection test results, parameterized query validation",
            "category": "Injection Testing"
        },
        {
            "id": "4.6",
            "requirement": "Cross-Site Scripting (XSS) testing is performed",
            "evidence": "XSS test scenarios, output encoding validation, CSP testing",
            "category": "Injection Testing"
        },
        {
            "id": "4.7",
            "requirement": "XXE and deserialization vulnerability testing is performed",
            "evidence": "XXE test cases, unsafe deserialization tests, XML/JSON validation",
            "category": "Injection Testing"
        },
        {
            "id": "4.8",
            "requirement": "API fuzzing is performed to discover unexpected behaviour",
            "evidence": "Fuzzing tool results (Burp Intruder, ffuf, RESTler), crash reports",
            "category": "API Testing"
        },
        {
            "id": "4.9",
            "requirement": "API rate limiting and abuse prevention are tested",
            "evidence": "Rate limit test evidence, DoS protection testing, throttling validation",
            "category": "API Testing"
        },
        {
            "id": "4.10",
            "requirement": "CORS configuration is validated for security",
            "evidence": "CORS test results, origin validation tests, preflight testing",
            "category": "API Testing"
        },
        {
            "id": "4.11",
            "requirement": "Security headers are validated in API responses",
            "evidence": "Header validation tests (HSTS, CSP, X-Frame-Options, etc.)",
            "category": "Application Testing"
        },
        {
            "id": "4.12",
            "requirement": "API error responses are tested for information disclosure",
            "evidence": "Error response security tests, stack trace prevention validation",
            "category": "API Testing"
        },
        {
            "id": "4.13",
            "requirement": "Staging/testing environment mirrors production security configuration",
            "evidence": "Environment parity documentation, configuration comparison",
            "category": "Test Environment"
        },
        {
            "id": "4.14",
            "requirement": "Test data is sanitized and does not include production data",
            "evidence": "Test data management policy, data anonymization procedures",
            "category": "Test Environment"
        },
        {
            "id": "4.15",
            "requirement": "API documentation is security-reviewed for sensitive information disclosure",
            "evidence": "API documentation review records, Swagger/OpenAPI security review",
            "category": "API Testing"
        },
        {
            "id": "4.16",
            "requirement": "GraphQL-specific security testing is performed (if applicable)",
            "evidence": "GraphQL introspection testing, query depth limiting, batching abuse tests",
            "category": "API Testing"
        },
        {
            "id": "4.17",
            "requirement": "WebSocket security is tested (if applicable)",
            "evidence": "WebSocket test scenarios, connection hijacking tests, message validation",
            "category": "Application Testing"
        },
        {
            "id": "4.18",
            "requirement": "API versioning security is validated across versions",
            "evidence": "Version security tests, deprecated endpoint validation, migration testing",
            "category": "API Testing"
        },
    ]


def get_domain5_requirements():
    """
    Domain 5: External Testing & Continuous Validation
    
    External perspectives find issues internal testing misses. Pentests and
    bug bounty programs provide reality checks on security posture.
    """
    return [
        {
            "id": "5.1",
            "requirement": "Annual penetration testing is performed for critical applications",
            "evidence": "Penetration test reports from last 12 months, test schedules",
            "category": "Penetration Testing"
        },
        {
            "id": "5.2",
            "requirement": "Penetration test scope covers all critical applications and APIs",
            "evidence": "Pentest scope documentation, application inventory",
            "category": "Penetration Testing"
        },
        {
            "id": "5.3",
            "requirement": "Penetration test findings are prioritised by risk and severity",
            "evidence": "Risk-ranked findings list, severity classification methodology",
            "category": "Penetration Testing"
        },
        {
            "id": "5.4",
            "requirement": "Critical pentest findings are remediated within 30 days",
            "evidence": "Remediation tracking system, SLA compliance metrics",
            "category": "Remediation"
        },
        {
            "id": "5.5",
            "requirement": "Penetration test retest validates that fixes are effective",
            "evidence": "Retest reports, fix validation evidence",
            "category": "Penetration Testing"
        },
        {
            "id": "5.6",
            "requirement": "Bug bounty or responsible disclosure program is active",
            "evidence": "Bug bounty platform documentation (HackerOne, Bugcrowd), program page",
            "category": "Bug Bounty"
        },
        {
            "id": "5.7",
            "requirement": "Vulnerability disclosure policy is publicly published",
            "evidence": "Published security.txt or vulnerability disclosure policy URL",
            "category": "Bug Bounty"
        },
        {
            "id": "5.8",
            "requirement": "External vulnerability submissions are triaged within 48 hours",
            "evidence": "Triage SLA metrics, response time tracking",
            "category": "Bug Bounty"
        },
        {
            "id": "5.9",
            "requirement": "Security researcher communication is handled professionally",
            "evidence": "Communication templates, researcher feedback, response examples",
            "category": "Bug Bounty"
        },
        {
            "id": "5.10",
            "requirement": "Bug bounty payments are tracked and processed timely",
            "evidence": "Payment records, payout tracking, bounty statistics",
            "category": "Bug Bounty"
        },
        {
            "id": "5.11",
            "requirement": "Security regression tests exist for all past critical vulnerabilities",
            "evidence": "Regression test suite, vulnerability-specific tests",
            "category": "Regression Testing"
        },
        {
            "id": "5.12",
            "requirement": "Security regression tests run automatically in CI/CD",
            "evidence": "CI/CD regression test execution logs, test automation",
            "category": "Regression Testing"
        },
        {
            "id": "5.13",
            "requirement": "Security testing metrics are tracked over time",
            "evidence": "Historical metrics dashboard, trend analysis",
            "category": "Metrics"
        },
        {
            "id": "5.14",
            "requirement": "Vulnerability trends are analysed to identify patterns",
            "evidence": "Vulnerability trend analysis reports, root cause analysis",
            "category": "Metrics"
        },
        {
            "id": "5.15",
            "requirement": "Testing effectiveness is measured with KPIs",
            "evidence": "Testing effectiveness KPIs, detection rate metrics",
            "category": "Metrics"
        },
        {
            "id": "5.16",
            "requirement": "False negative analysis is performed after security incidents",
            "evidence": "Post-incident analysis reports, testing gap assessments",
            "category": "Continuous Improvement"
        },
        {
            "id": "5.17",
            "requirement": "Testing gaps identified in incidents are addressed systematically",
            "evidence": "Gap remediation plans, testing enhancement initiatives",
            "category": "Continuous Improvement"
        },
        {
            "id": "5.18",
            "requirement": "Continuous improvement process drives testing program evolution",
            "evidence": "Improvement backlog, testing maturity roadmap",
            "category": "Continuous Improvement"
        },
    ]

# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET CREATION
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each assessment domain sheet sequentially.', '2. Focus on EFFECTIVENESS, not just process existence.', '3. For each requirement, select Implementation Status from the dropdown menu.', '4. Provide evidence references documenting WHERE evidence can be found.', '5. Sample actual code reviews to validate quality (not just process documentation).', '6. Review the Summary Dashboard to see overall compliance metrics.', '7. Document gaps with specific, actionable remediation plans.', '8. Obtain approvals from both Security and Engineering leadership.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_domain_sheet(ws, domain_name, requirements, styles):
    """Create standardised assessment domain sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Code Review & Testing Assessment: {domain_name}".upper()
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO 27001:2022 | Control A.8.28 | {domain_name}"
    ws["A2"].font = Font(italic=True, size=10, color="003366", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["ID", "Requirement", "Implementation Status", "Evidence Reference", "Comments"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        apply_style(cell, styles["column_header"])
    
    ws.row_dimensions[3].height = 30
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40

    current_category = None
    row = 4
    
    for req in requirements:
        if "category" in req and req["category"] != current_category:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"▶ {req['category']}"
            ws[f"A{row}"].font = Font(bold=True, size=11, color="4472C4")
            ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            current_category = req["category"]
        
        ws[f"A{row}"] = req["id"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        ws[f"B{row}"] = req["requirement"]
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        validations["implementation_status"].add(ws[f"C{row}"])
        
        ws[f"D{row}"] = req.get("evidence", "")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"D{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[f"D{row}"].font = Font(size=9, italic=True, color="666666")
        
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"E{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        
        row += 1

    # Pad to 50 FFFFCC data rows for QA compliance (DS-005)
    min_data_rows = 50
    data_rows_written = len(requirements)
    if data_rows_written < min_data_rows:
        thin = Side(style="thin")
        for pad_row in range(row, row + (min_data_rows - data_rows_written)):
            for col in ["A", "B", "C", "D", "E"]:
                cell = ws[f"{col}{pad_row}"]
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws[f"C{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"C{pad_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            validations["implementation_status"].add(ws[f"C{pad_row}"])
            ws[f"D{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"D{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws[f"E{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"E{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Hide unused columns (F onwards) to prevent empty column display
    for col_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    return ws


def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard matching A.8.33 gold standard with TABLE 1, 2, 3."""
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (italic, 003366 font, no fill)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 Annex A.8.28 — Secure Coding | Code Review & Testing Assessment"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4: TABLE 1 banner (003366, merged A4:G4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=4, column=col).border = border

    # Row 5: Column headers (D9D9D9 grey, black text)
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows (4-8) - one per assessment area
    area_configs = [
        ("Code Review Process", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Security Champion Review", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Unit Integration Testing", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("API Application Testing", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("External Testing Validation", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
    ]

    for i, (area_name, status_col, status_values) in enumerate(area_configs):
        row = 6 + i

        # Column A: Area name
        ws.cell(row=row, column=1, value=area_name).border = border

        # Column B: Total Items (COUNTA of status column)
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTA('{area_name}'!{status_col}4:{status_col}100)"
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal="center")

        # Column C: Compliant (✅ Implemented)
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[0]}")'
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal="center")

        # Column D: Partial (⚠️ Partially Implemented)
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[1]}")'
        cell_d.border = border
        cell_d.alignment = Alignment(horizontal="center")

        # Column E: Non-Compliant (❌ Not Implemented)
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[2]}")'
        cell_e.border = border
        cell_e.alignment = Alignment(horizontal="center")

        # Column F: N/A count
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"N/A")'
        cell_f.border = border
        cell_f.alignment = Alignment(horizontal="center")

        # Column G: Compliance % (Compliant / (Total - N/A))
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f'=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))'
        cell_g.number_format = "0.0%"
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center")

    # TOTAL row (after data rows + 2 buffer rows)
    total_row = 6 + len(area_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1).border = border
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    # Compliance % formula
    cell = ws.cell(row=total_row, column=7)
    cell.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    cell.number_format = "0.0%"
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    # TABLE 2: Key Metrics
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{metrics_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=metrics_start, column=col).border = border

    # TABLE 2 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        hcell = ws.cell(row=metrics_start + 1, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center")

    # Metrics exploiting DVs from all 5 sheets
    metrics = [
        # Security Requirements & Design
        ("Total Requirements Tracked", "=COUNTA('Code Review Process'!B5:B100)"),
        ("Security Requirements Documented", '=COUNTIF(\'Code Review Process\'!C5:C100,"✅ Implemented")'),
        ("Threat Modeling Coverage", '=COUNTIF(\'Code Review Process\'!C5:C100,"✅ Implemented")/COUNTA(\'Code Review Process\'!B5:B100)'),
        # Development Environment
        ("Total Dev Environment Controls", "=COUNTA('Security Champion Review'!B5:B100)"),
        ("Workstations with Security Tools", '=COUNTIF(\'Security Champion Review\'!C5:C100,"✅ Implemented")'),
        ("Repository Security Controls", '=COUNTIF(\'Security Champion Review\'!C5:C100,"✅ Implemented")'),
        # Build & Deployment
        ("Total Pipeline Controls", "=COUNTA('Unit Integration Testing'!B5:B100)"),
        ("Security Scans Integrated", '=COUNTIF(\'Unit Integration Testing\'!C5:C100,"✅ Implemented")'),
        ("Deployment Gates Active", '=COUNTIF(\'Unit Integration Testing\'!C5:C100,"✅ Implemented")'),
        # Security Testing
        ("Total Testing Controls", "=COUNTA('API Application Testing'!B5:B100)"),
        ("SAST/DAST Integration", '=COUNTIF(\'API Application Testing\'!C5:C100,"✅ Implemented")'),
        ("Penetration Testing Coverage", '=COUNTIF(\'API Application Testing\'!C5:C100,"✅ Implemented")'),
        # Release & Change Management
        ("Total Release Controls", "=COUNTA('External Testing Validation'!B5:B100)"),
        ("Approval Workflows Active", '=COUNTIF(\'External Testing Validation\'!C5:C100,"✅ Implemented")'),
        ("Change Management Compliance", '=COUNTIF(\'External Testing Validation\'!C5:C100,"✅ Implemented")/COUNTA(\'External Testing Validation\'!B5:B100)'),
    ]

    r = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=r, column=1, value=metric).border = border
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10)
        cell_val = ws.cell(row=r, column=2, value=formula)
        cell_val.border = border
        cell_val.font = Font(name="Calibri", size=10, color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # 2 buffer rows after TABLE 2 metrics (white, bordered)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # TABLE 3: Critical Findings (red banner)
    crit_start = r + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    # TABLE 3 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Critical findings based on A.8.28 secure coding requirements
    findings = [
        ("Code Review Process", "Requirements without security acceptance criteria", '=COUNTIF(\'Code Review Process\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Security Champion Review", "Developer workstations without EDR/security tools", '=COUNTIF(\'Security Champion Review\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Unit Integration Testing", "Pipelines without SAST/SCA scanning", '=COUNTIF(\'Unit Integration Testing\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("API Application Testing", "Production code without security testing", '=COUNTIF(\'API Application Testing\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("External Testing Validation", "Releases without security approval", '=COUNTIF(\'External Testing Validation\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Code Review Process", "Missing threat modeling for high-risk apps", '=COUNTIF(\'Code Review Process\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Security Champion Review", "Repositories without MFA enforcement", '=COUNTIF(\'Security Champion Review\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Unit Integration Testing", "Security gates not enforced", '=COUNTIF(\'Unit Integration Testing\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("API Application Testing", "Incomplete DAST coverage", '=COUNTIF(\'API Application Testing\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("External Testing Validation", "Emergency changes without post-review", '=COUNTIF(\'External Testing Validation\'!C5:C100,"⚠️ Partially Implemented")', "Medium", "Plan"),
    ]

    r3 = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        ws.cell(row=r3, column=1, value=cat)
        ws.cell(row=r3, column=2, value=finding)
        count_cell = ws.cell(row=r3, column=3, value=formula)
        count_cell.alignment = Alignment(horizontal="center")
        count_cell.font = Font(name="Calibri", size=10, color="000000")
        ws.cell(row=r3, column=4, value=severity)
        ws.cell(row=r3, column=5, value=action)
        r3 += 1

    # 2 buffer rows after TABLE 3 (FFFFCC)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        r3 += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"  # Freeze below subtitle + banner + headers

    return ws


# ============================================================================
# SECTION 7: EVIDENCE REGISTER CREATION
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        ("Evidence ID", 15),
        ("Assessment Area", 25),
        ("Evidence Type", 22),
        ("Description", 40),
        ("Location/Path", 45),
        ("Date Collected", 16),
        ("Collected By", 20),
        ("Verification Status", 22),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border_thin
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Data rows EV-001 to EV-100
    dv_etype = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=False
    )
    dv_vstatus = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False
    )
    ws.add_data_validation(dv_etype)
    ws.add_data_validation(dv_vstatus)

    # MAX-001/MAX-002 fix: 1 grey sample + 100 FFFFCC empty rows
    # Sample row (row 5) with F2F2F2 grey background (like A.8.33 reference)
    sample_data = [
        "EV-001",
        "Test Data Inventory",
        "Configuration file",
        "Data masking configuration for UAT environment",
        "\\fileserver\\isms\\evidence\\test-data\\uat-masking-config.json",
        "2024-03-01",
        "Sarah Mitchell",
        "Verified"
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=5, column=col_num, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = border_thin

    # 100 empty rows (rows 6-105) - FFFFCC yellow
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
        dv_etype.add(ws.cell(row=r, column=3))
        dv_vstatus.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"
    return ws


def create_gap_analysis_sheet(ws, styles):
    """Create Gap Analysis sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells("A1:J1")
    ws["A1"] = "GAP ANALYSIS & REMEDIATION PLAN"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all 'Not Implemented' and 'Partially Implemented' requirements. Assign owners, set target dates, and track remediation progress."
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Gap ID", "Domain", "Requirement ID", "Requirement Description", "Current State", "Target State", "Priority", "Owner", "Target Date", "Status"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    widths = {"A": 10, "B": 22, "C": 12, "D": 35, "E": 25, "F": 25, "G": 12, "H": 20, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    validations["priority"].add("G4:G200")
    validations["gap_status"].add("J4:J200")

    # MAX-003 fix: 1 sample row (F2F2F2 grey) + 50 empty rows (FFFFCC yellow) per Option B standard
    thin = Side(style="thin")

    # Sample row (row 4) with realistic example data
    sample_data = [
        "GAP-001",
        "Code Review Process",
        "1.3",
        "Security-focused code review checklist used",
        "Generic code review without security focus",
        "OWASP Code Review Guide checklist integrated",
        "Critical",
        "Security Champion",
        "15.02.2025",
        "Open"
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col_num, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 50 empty rows (rows 5-54) - FFFFCC yellow
    for gap_row in range(5, 55):
        for col in range(1, 11):
            cell = ws.cell(row=gap_row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Hide unused columns (K onwards) to prevent empty column display in Gap Analysis
    for col_letter in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    return ws


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=1, column=col).border = border_thin
    ws.row_dimensions[1].height = 35

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=2, column=col).border = border_thin

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=row, column=col).border = border_thin

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.28.3 - Code Review & Testing Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        # Apply borders to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = border_thin
            if value == "":
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Status dropdown on Assessment Status
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Apply borders to banner (A:E)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_thin
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            # Apply borders and fill to all cells in merged range (B:E)
            for col in range(2, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=row, column=col).border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all cells in merged range (B:E)
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border_thin

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to banner (A:E)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border_thin

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

    return ws



def main():
    """
    Generate Code Review & Testing Assessment workbook.
    
    Reviews and tests are the last lines of defense before production.
    This assessment helps determine if those defenses actually work.
    """
    logger.info("=" * 80)
    logger.info(" " * 12 + "ISMS Control 8.28.3 - Code Review & Testing Assessment Generator")
    logger.info("=" * 80)
    logger.info("")
    
    try:
        logger.info("[N] [1/10] Creating workbook structure...")
        wb = create_workbook()
        styles = _STYLES
        logger.info("     ✓ Workbook initialized with 10 sheets")

        logger.info("[F] [2/10] Creating Instructions sheet...")
        ws_instructions = wb["Instructions & Legend"]
        ws_instructions.sheet_view.showGridLines = False
        create_instructions_sheet(ws_instructions)
        logger.info("     ✓ Instructions complete (focus: effectiveness, not theater)")

        logger.info("[CHART] [3/10] Creating Domain 1: Code Review Process & Governance...")
        ws_domain1 = wb["Code Review Process"]
        ws_domain1.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain1, "Code Review Process & Governance", get_domain1_requirements(), styles)
        logger.info("     ✓ 18 requirements (Review Policy, Quality, Metrics)")

        logger.info("[CHART] [4/10] Creating Domain 2: Security Champion Review & Architecture...")
        ws_domain2 = wb["Security Champion Review"]
        ws_domain2.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain2, "Security Champion Review & Architecture", get_domain2_requirements(), styles)
        logger.info("     ✓ 18 requirements (Champion Program, Threat Modeling, Design Review)")

        logger.info("[CHART] [5/10] Creating Domain 3: Security Testing - Unit & Integration...")
        ws_domain3 = wb["Unit Integration Testing"]
        ws_domain3.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain3, "Security Testing - Unit & Integration", get_domain3_requirements(), styles)
        logger.info("     ✓ 18 requirements (Unit Tests, Integration Tests, Test Coverage)")

        logger.info("[CHART] [6/10] Creating Domain 4: API & Application Security Testing...")
        ws_domain4 = wb["API Application Testing"]
        ws_domain4.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain4, "API & Application Security Testing", get_domain4_requirements(), styles)
        logger.info("     ✓ 18 requirements (API Testing, Injection Testing, Fuzzing)")

        logger.info("[CHART] [7/10] Creating Domain 5: External Testing & Continuous Validation...")
        ws_domain5 = wb["External Testing Validation"]
        ws_domain5.sheet_view.showGridLines = False
        create_domain_sheet(ws_domain5, "External Testing & Continuous Validation", get_domain5_requirements(), styles)
        logger.info("     ✓ 18 requirements (Pentesting, Bug Bounty, Regression Tests)")

        logger.info("[TREND] [8/10] Creating Summary Dashboard...")
        ws_summary = wb["Summary Dashboard"]
        ws_summary.sheet_view.showGridLines = False
        create_summary_dashboard_sheet(ws_summary, styles)
        logger.info("     ✓ Executive summary with effectiveness insights")

        logger.info("[9/10] Creating Evidence Register...")
        ws_evidence = wb["Evidence Register"]
        ws_evidence.sheet_view.showGridLines = False
        create_evidence_register(ws_evidence, styles)
        logger.info("     ✓ Evidence tracking")

        logger.info("\u26A0\uFE0F  [10/10] Creating Gap Analysis & Approval sheets...")
        ws_gap = wb["Gap Analysis"]
        ws_gap.sheet_view.showGridLines = False
        create_gap_analysis_sheet(ws_gap, styles)
        
        ws_approval = wb["Approval Sign-Off"]
        ws_approval.sheet_view.showGridLines = False
        create_approval_sheet(ws_approval, styles)
        logger.info("     ✓ Gap tracking and approval workflow")

        logger.info("")
        logger.info("[S] Saving workbook...")
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.8.28.3_Code_Review_Testing_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info("")
        logger.info("=" * 80)
        logger.info("\u2705 SUCCESS: Code Review & Testing Assessment workbook generated!")
        logger.info("=" * 80)
        logger.info("")
        logger.info(f"[D] File: {filename}")
        logger.info(f"[CHART] Total Requirements: 90 (18 per domain × 5 domains)")
        logger.info("")
        logger.info("[PIN] Assessment Domains:")
        logger.info("   1. Code Review Process & Governance    (18 requirements)")
        logger.info("   2. Security Champion Review            (18 requirements)")
        logger.info("   3. Unit & Integration Testing          (18 requirements)")
        logger.info("   4. API & Application Testing           (18 requirements)")
        logger.info("   5. External Testing & Validation       (18 requirements)")
        logger.info("")
        logger.info("[TIP] Key Focus: Does it catch vulnerabilities? Not just 'does the process exist?'")
        logger.info("")
        logger.info("\u26A0\uFE0F  Remember:")
        logger.info("   \u2022 'LGTM' without security analysis is not a review")
        logger.info("   \u2022 Tests that always pass don't test security")
        logger.info("   \u2022 Avoid Review Theater and Test Theater")
        logger.info("   \u2022 Focus on EFFECTIVENESS, not just process existence")
        logger.info("")
        logger.info("[CHART] Effectiveness Questions to Answer:")
        logger.info("   \u2022 What % of vulnerabilities are caught in review vs. testing vs. production?")
        logger.info("   \u2022 Are pentests finding issues internal testing should have caught?")
        logger.info("   \u2022 How many security regressions occur?")
        logger.info("")
        logger.info("=" * 80)
        
        return 0

    except Exception as e:
        logger.info("")
        logger.info("=" * 80)
        logger.error("\u274C ERROR: Failed to generate workbook")
        logger.info("=" * 80)
        logger.error(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
