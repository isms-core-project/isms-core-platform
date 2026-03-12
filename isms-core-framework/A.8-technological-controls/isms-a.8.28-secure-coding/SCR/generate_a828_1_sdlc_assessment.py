#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.28.1 - SDLC Integration Assessment
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Assessment Domain 1 of 4: Secure Development Lifecycle Integration

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific development environment, SDLC methodology, and
assessment requirements.

Key customisation areas:
1. SDLC methodology and phases (Agile, Waterfall, DevOps - match your process)
2. Security gate criteria and thresholds (adapt to your risk tolerance)
3. Development tools and platforms (specific to your toolchain)
4. Project types and classifications (based on your portfolio structure)
5. Compliance criteria and scoring (aligned with your regulatory requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMISE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
security integration across all phases of the Software Development Lifecycle.

**Purpose:**
Enables systematic assessment of SDLC security integration against ISO 27001:2022
Control A.8.28 requirements, supporting evidence-based validation of secure
development practices from requirements through deployment.

**Assessment Scope:**
- Security requirements definition and threat modeling processes
- Secure design review and architecture security validation
- Security testing integration (SAST, DAST, SCA, penetration testing)
- Security gate implementation and enforcement mechanisms
- Deployment security controls and production hardening
- Post-deployment monitoring and incident response integration
- Developer security training and awareness programs
- Security champion and SME involvement in SDLC
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and SDLC security standards
2. Requirements Phase - Security requirements and threat modeling assessment
3. Design Phase - Secure architecture and design review practices
4. Implementation Phase - Coding standards and secure development practices
5. Testing Phase - Security testing integration and coverage
6. Deployment Phase - Production security controls and hardening
7. Maintenance Phase - Patching, monitoring, and incident response
8. Training & Awareness - Developer security competency assessment
9. Security Gates - Gate effectiveness and enforcement evaluation
10. Gap Analysis - Non-compliant SDLC phases and remediation requirements
11. Evidence Register - Audit evidence tracking and documentation
12. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with SDLC methodology dropdown lists
- Conditional formatting for security gate compliance status
- Automated gap identification for missing security activities
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with project management and CI/CD tool outputs

**Integration:**
This assessment feeds into the A.8.28.5 Compliance Dashboard, which
consolidates data from all four assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a828_1_sdlc_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a828_1_sdlc_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a828_1_sdlc_assessment.py --date 20250124

Output:
    File: ISMS_A_8_28_1_SDLC_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise SDLC security standards to match your methodology
    2. Inventory all active development projects and their SDLC phases
    3. Complete assessments for each project or development stream
    4. Validate security gate effectiveness against defined criteria
    5. Review developer security training completion and competency
    6. Conduct gap analysis for missing or ineffective security activities
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (gate approvals, test reports, training records)
    9. Obtain stakeholder approvals (Dev Manager, AppSec Lead, CISO)
    10. Feed results into A.8.28.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    1 of 4 (SDLC Integration)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.1: SDLC Integration Implementation Guide
    - ISMS-IMP-A.8.28.2: Standards & Tools Assessment (Domain 2)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Assessment (Domain 3)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Assessment (Domain 4)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a828_2_standards_tools.py
    - generate_a828_3_code_review_testing.py
    - generate_a828_4_third_party_oss.py
    - generate_a828_5_compliance_dashboard.py
    - normalize_assessment_files_a828.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.28.1 specification
    - Supports comprehensive SDLC security integration evaluation
    - Integrated with A.8.28.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**SDLC Methodology Adaptation:**
This assessment framework is methodology-agnostic by design. Customise phase
names, security activities, and gate criteria to match your organisation's
specific SDLC approach (Agile/Scrum, Waterfall, DevOps, DevSecOps, SAFe, etc.).

**Security Gate Philosophy:**
Security gates should be enabling, not blocking. Focus assessment on whether
gates effectively identify and manage risk, not whether they create bottlenecks.
Overly rigid gates that teams bypass are worse than no gates at all.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of security gate approvals, testing results,
and training records.

**Data Protection:**
Assessment workbooks contain sensitive development information including:
- Project names, timelines, and architectures
- Vulnerability information and security gaps
- Development team structures and resource allocation
- Security testing results and penetration test findings

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check security gate effectiveness and compliance rates
- Semi-annually: Update security activity requirements for new threats
- Annually: Complete reassessment of all active development projects
- Ad-hoc: When SDLC methodology changes or new security requirements emerge

**Quality Assurance:**
Have application security SMEs, development managers, and security champions
validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
Ensure SDLC security standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 secure SDLC requirements (Requirement 6)
- Healthcare: HIPAA application security and risk analysis standards
- Finance: Regional banking application security requirements
- Government: Jurisdiction-specific secure development mandates (e.g., NIST SSDF)

Customise assessment criteria to include regulatory-specific requirements.

**Integration with Development Tools:**
Where possible, integrate assessment data with:
- Project management systems (Jira, Azure DevOps, etc.)
- CI/CD pipeline metrics (Jenkins, GitLab, GitHub Actions)
- Security testing tool outputs (SAST, DAST, SCA results)
- Training platforms (completion tracking, competency scores)

Automated tool integration reduces manual effort and improves accuracy.

**Don't Fool Yourself - Feynman Principle:**
Assess actual SDLC practice, not documented procedures. Just because a security
gate exists on paper doesn't mean it's effective. Look for evidence of:
- Actual gate approvals (not rubber-stamping)
- Real security defects found and fixed
- Developers actually trained (not just registered for courses)
- Threat models that influenced design (not checkbox exercises)

If you can't find evidence it's working, it probably isn't.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.28.1"
WORKBOOK_NAME = "Secure Development Lifecycle Integration"
CONTROL_ID = "A.8.28"
CONTROL_NAME = "Secure Coding"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


import sys

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.28.1 specification
    # 10 sheets total: Instructions + 5 assessment domains + 4 supporting sheets
    sheets = [
        "Instructions & Legend",
        "Security Requirements Design",
        "Development Environment",
        "Build Deployment Pipeline",
        "Security Testing Integration",
        "Release Change Management",
        "Evidence Register",
        "Gap Analysis",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    
    Following Feynman's principle: "The first principle is that you must not 
    fool yourself—and you are the easiest person to fool." We don't fool Excel
    with shared style objects that cause repair warnings.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_implemented": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notimplemented": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        },
        "priority_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "priority_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "priority_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "priority_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    
    This is cargo cult done right: we follow the ritual of creating new objects
    not because we worship Excel's quirks, but because we understand why
    shared style objects cause corruption.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    
    Data validations ensure consistency and reduce user error - this is
    not cargo cult, it's user interface design.
    """
    validations = {
        # Implementation status dropdowns
        'implementation_status': DataValidation(
            type="list",
            formula1='"✅ Implemented,⚠️ Partially Implemented,❌ Not Implemented,N/A"',
            allow_blank=False
        ),
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Unknown"',
            allow_blank=False
        ),
        'yes_no_planned': DataValidation(
            type="list",
            formula1='"Yes,No,Planned,N/A"',
            allow_blank=False
        ),
        
        # Priority and risk levels
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Negligible"',
            allow_blank=False
        ),
        
        # Gap analysis statuses
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Closed,Deferred"',
            allow_blank=False
        ),
        
        # Evidence types
        'evidence_type': DataValidation(
            type="list",
            formula1='"Document,Screenshot,Report,Configuration,URL,Log File,Diagram,Policy,Procedure,Test Result,Scan Result,Other"',
            allow_blank=False
        ),
        
        # Verification status
        'verification_status': DataValidation(
            type="list",
            formula1='"Verified,Pending,Not Verified,N/A"',
            allow_blank=False
        ),
        
        # Approval decision
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
            allow_blank=False
        ),
        
        # SDLC phase
        'sdlc_phase': DataValidation(
            type="list",
            formula1='"Requirements,Design,Development,Testing,Deployment,Maintenance"',
            allow_blank=False
        ),
        
        # Tool category
        'tool_category': DataValidation(
            type="list",
            formula1='"SAST,SCA,DAST,Secret Scanner,Container Scanner,IaC Scanner,IDE Plugin,Other"',
            allow_blank=False
        ),
        
        # Testing type
        'testing_type': DataValidation(
            type="list",
            formula1='"Unit,Integration,API,Security,Penetration,Performance,Regression"',
            allow_blank=False
        ),
        
        # Frequency
        'frequency': DataValidation(
            type="list",
            formula1='"Continuous,Daily,Weekly,Monthly,Quarterly,On-Demand"',
            allow_blank=False
        ),
        
        # Maturity level
        'maturity_level': DataValidation(
            type="list",
            formula1='"Initial,Repeatable,Defined,Managed,Optimizing"',
            allow_blank=False
        ),
        
        # Training status
        'training_status': DataValidation(
            type="list",
            formula1='"Completed,In Progress,Not Started,N/A"',
            allow_blank=False
        ),
        
        # Deployment environment
        'environment': DataValidation(
            type="list",
            formula1='"Development,Testing,Staging,Production,All"',
            allow_blank=False
        ),
        
        # Automation level
        'automation_level': DataValidation(
            type="list",
            formula1='"Fully Automated,Partially Automated,Manual,Not Applicable"',
            allow_blank=False
        ),
        
        # Compliance level
        'compliance_level': DataValidation(
            type="list",
            formula1='"Fully Compliant,Mostly Compliant,Partially Compliant,Non-Compliant,N/A"',
            allow_blank=False
        ),
    }

    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: ASSESSMENT DOMAIN DATA DEFINITIONS
# ============================================================================

def get_domain1_requirements():
    """
    Domain 1: Security Requirements & Design
    
    This domain assesses whether security is "baked in" from the start,
    not bolted on later. As Feynman might say: "If you can't specify 
    security requirements upfront, you're just fooling yourself about 
    having a secure development process."
    """
    return [
        {
            "id": "1.1",
            "requirement": "Security requirements are documented for all new development projects",
            "evidence": "Requirements documents, user stories with security acceptance criteria, requirements management system records",
            "category": "Requirements Definition"
        },
        {
            "id": "1.2",
            "requirement": "Threat modeling is performed for high-risk applications before implementation",
            "evidence": "Threat model documents, STRIDE analysis, attack trees, threat modeling session notes",
            "category": "Requirements Definition"
        },
        {
            "id": "1.3",
            "requirement": "Security design reviews are conducted before implementation begins",
            "evidence": "Architecture review meeting notes, approved design documents, security sign-off records",
            "category": "Design Review"
        },
        {
            "id": "1.4",
            "requirement": "Authentication and authorisation requirements are defined upfront for all applications",
            "evidence": "Security specifications, access control matrices, identity management requirements",
            "category": "Requirements Definition"
        },
        {
            "id": "1.5",
            "requirement": "Data classification is performed for all data handled by applications",
            "evidence": "Data classification register, data flow diagrams, data handling requirements",
            "category": "Requirements Definition"
        },
        {
            "id": "1.6",
            "requirement": "Cryptographic requirements are specified when sensitive data is involved",
            "evidence": "Crypto standards documentation, algorithm approval records, key management specs",
            "category": "Requirements Definition"
        },
        {
            "id": "1.7",
            "requirement": "Security requirements are tracked throughout the entire SDLC",
            "evidence": "Requirement traceability matrix, ticket system reports, requirements coverage metrics",
            "category": "Requirements Management"
        },
        {
            "id": "1.8",
            "requirement": "Privacy requirements (GDPR, CCPA, etc.) are incorporated into design phase",
            "evidence": "Privacy impact assessments, data protection by design documentation, consent mechanisms",
            "category": "Requirements Definition"
        },
        {
            "id": "1.9",
            "requirement": "Secure coding standards are referenced in all project documentation",
            "evidence": "Project wikis, coding guidelines links, onboarding documentation",
            "category": "Design Review"
        },
        {
            "id": "1.10",
            "requirement": "Security acceptance criteria are defined for all user stories and features",
            "evidence": "JIRA/Azure DevOps stories with security DoD, acceptance criteria templates",
            "category": "Requirements Definition"
        },
        {
            "id": "1.11",
            "requirement": "API security requirements are documented for all APIs (internal and external)",
            "evidence": "API security specifications, OpenAPI security schemes, API gateway policies",
            "category": "Requirements Definition"
        },
        {
            "id": "1.12",
            "requirement": "Third-party integration security is assessed during design phase",
            "evidence": "Vendor security assessments, integration security reviews, API security analysis",
            "category": "Design Review"
        },
        {
            "id": "1.13",
            "requirement": "Error handling and logging requirements are specified in design",
            "evidence": "Logging standards, error handling patterns documentation, log retention policies",
            "category": "Design Review"
        },
        {
            "id": "1.14",
            "requirement": "Security testing requirements are defined before development starts",
            "evidence": "Test plans, security test case specifications, testing strategy documents",
            "category": "Requirements Definition"
        },
        {
            "id": "1.15",
            "requirement": "Compliance requirements are mapped to technical controls in design",
            "evidence": "Compliance mapping documents, control implementation matrix, regulatory requirements docs",
            "category": "Requirements Definition"
        },
        {
            "id": "1.16",
            "requirement": "Security Champions are involved in requirements and design review",
            "evidence": "Review meeting attendance, Security Champion feedback records, design review checklists",
            "category": "Design Review"
        },
        {
            "id": "1.17",
            "requirement": "Secure design patterns are documented and promoted for common scenarios",
            "evidence": "Design pattern library, secure architecture templates, reference architectures",
            "category": "Design Review"
        },
        {
            "id": "1.18",
            "requirement": "Security requirements are risk-based and proportional to data sensitivity",
            "evidence": "Risk assessment documentation, security control selection justification",
            "category": "Requirements Definition"
        },
    ]


def get_domain2_requirements():
    """
    Domain 2: Development Environment & Tooling
    
    The development environment is where vulnerabilities are often introduced.
    Proper tooling and configuration prevents many security issues before
    they reach production.
    """
    return [
        {
            "id": "2.1",
            "requirement": "Development environments are logically isolated from production systems",
            "evidence": "Network diagrams, firewall rules, network segmentation documentation",
            "category": "Environment Security"
        },
        {
            "id": "2.2",
            "requirement": "Developer workstations have mandatory security tooling installed (EDR, DLP, etc.)",
            "evidence": "EDR deployment reports, endpoint security compliance dashboards, security agent status",
            "category": "Workstation Security"
        },
        {
            "id": "2.3",
            "requirement": "IDE security plugins are available and promoted to all developers",
            "evidence": "IDE plugin catalog, installation instructions, adoption metrics",
            "category": "Development Tools"
        },
        {
            "id": "2.4",
            "requirement": "Pre-commit hooks prevent secret commits to version control",
            "evidence": "Git hooks configuration, secret scanner setup (git-secrets, detect-secrets), violation reports",
            "category": "Development Tools"
        },
        {
            "id": "2.5",
            "requirement": "Code repositories require authentication and authorisation (SSO/MFA)",
            "evidence": "GitHub/GitLab access logs, SSO configuration, MFA enforcement policies",
            "category": "Repository Security"
        },
        {
            "id": "2.6",
            "requirement": "Branch protection rules enforce mandatory peer review for production branches",
            "evidence": "Repository settings screenshots, branch protection rules configuration, merge policies",
            "category": "Repository Security"
        },
        {
            "id": "2.7",
            "requirement": "Development and testing use anonymized or synthetic data (not production data)",
            "evidence": "Data masking procedures, test data generation tools, data usage policies",
            "category": "Environment Security"
        },
        {
            "id": "2.8",
            "requirement": "Secrets management solution is available and documented for developers",
            "evidence": "Vault/secrets manager documentation, developer usage guidelines, secrets rotation policies",
            "category": "Development Tools"
        },
        {
            "id": "2.9",
            "requirement": "Local development environments use HTTPS/TLS where applicable",
            "evidence": "Development proxy configs, local cert management, TLS enforcement policies",
            "category": "Environment Security"
        },
        {
            "id": "2.10",
            "requirement": "Dependency management tools scan for vulnerabilities in developer environments",
            "evidence": "npm audit, pip-audit, Dependabot integration, vulnerability scan results",
            "category": "Development Tools"
        },
        {
            "id": "2.11",
            "requirement": "Development environment hardening guidelines exist and are enforced",
            "evidence": "Developer workstation security baseline, configuration management records",
            "category": "Workstation Security"
        },
        {
            "id": "2.12",
            "requirement": "Multi-factor authentication (MFA) is enforced for all code repository access",
            "evidence": "MFA enforcement policies, authentication logs, MFA adoption metrics",
            "category": "Repository Security"
        },
        {
            "id": "2.13",
            "requirement": "Development tools and IDEs are kept up-to-date with security patches",
            "evidence": "Tool version inventory, patch management records, software update policies",
            "category": "Development Tools"
        },
        {
            "id": "2.14",
            "requirement": "Secure coding training resources are accessible to all developers",
            "evidence": "Training portal links, OWASP resources, internal wiki, training completion records",
            "category": "Training & Awareness"
        },
        {
            "id": "2.15",
            "requirement": "Development containers/VMs use secure, hardened base images",
            "evidence": "Approved image catalog, image scanning results, base image security standards",
            "category": "Environment Security"
        },
        {
            "id": "2.16",
            "requirement": "Code repository activity is logged and monitored for anomalies",
            "evidence": "Repository audit logs, SIEM integration, anomaly detection alerts",
            "category": "Repository Security"
        },
        {
            "id": "2.17",
            "requirement": "Personal access tokens and API keys have expiration policies",
            "evidence": "Token lifecycle policies, key rotation procedures, token expiration reports",
            "category": "Development Tools"
        },
        {
            "id": "2.18",
            "requirement": "Developers have access to security documentation and secure code examples",
            "evidence": "Internal security wiki, code example repository, secure coding cheat sheets",
            "category": "Training & Awareness"
        },
    ]


def get_domain3_requirements():
    """
    Domain 3: Build & Deployment Pipeline
    
    The CI/CD pipeline is the assembly line where security checks happen
    automatically. If it's not in the pipeline, it doesn't happen consistently.
    """
    return [
        {
            "id": "3.1",
            "requirement": "CI/CD pipelines enforce automated security checks on every build",
            "evidence": "Pipeline configuration files (Jenkinsfile, .gitlab-ci.yml), build logs showing security scans",
            "category": "Pipeline Security"
        },
        {
            "id": "3.2",
            "requirement": "SAST (Static Application Security Testing) tools are integrated into build pipeline",
            "evidence": "SAST scan results, tool integration config (SonarQube, Semgrep, Checkmarx)",
            "category": "Security Scanning"
        },
        {
            "id": "3.3",
            "requirement": "SCA (Software Composition Analysis) tools scan dependencies during every build",
            "evidence": "SCA scan reports (Snyk, WhiteSource, Dependency-Check), vulnerability alerts",
            "category": "Security Scanning"
        },
        {
            "id": "3.4",
            "requirement": "Secret scanning runs automatically on every commit",
            "evidence": "Secret scanner logs (TruffleHog, GitGuardian), detected secret alerts, remediation records",
            "category": "Security Scanning"
        },
        {
            "id": "3.5",
            "requirement": "Build artifacts are signed and signature verification is enforced",
            "evidence": "Code signing certificates, artifact signatures, signature verification logs",
            "category": "Artifact Security"
        },
        {
            "id": "3.6",
            "requirement": "Container images are scanned for vulnerabilities before deployment",
            "evidence": "Trivy/Clair/Anchore scan results, image vulnerability reports, scan policies",
            "category": "Security Scanning"
        },
        {
            "id": "3.7",
            "requirement": "Infrastructure-as-Code (IaC) is scanned for security misconfigurations",
            "evidence": "Checkov/tfsec/Terrascan results, IaC security scan reports, policy violations",
            "category": "Security Scanning"
        },
        {
            "id": "3.8",
            "requirement": "Build pipelines fail on critical or high severity vulnerabilities",
            "evidence": "Pipeline failure logs, quality gate configuration, vulnerability threshold policies",
            "category": "Pipeline Security"
        },
        {
            "id": "3.9",
            "requirement": "Deployment to production requires explicit approval",
            "evidence": "Approval workflows, deployment gate records, approval audit logs",
            "category": "Deployment Controls"
        },
        {
            "id": "3.10",
            "requirement": "Pipeline service accounts use least-privilege access principles",
            "evidence": "Service account permissions, IAM policies, privilege reviews",
            "category": "Pipeline Security"
        },
        {
            "id": "3.11",
            "requirement": "Build environments are ephemeral and disposable (not persistent)",
            "evidence": "Build agent lifecycle documentation, container/VM configs, environment refresh logs",
            "category": "Pipeline Security"
        },
        {
            "id": "3.12",
            "requirement": "Pipeline configurations are version-controlled and peer-reviewed",
            "evidence": "CI/CD config files in git, change history, pipeline review records",
            "category": "Pipeline Security"
        },
        {
            "id": "3.13",
            "requirement": "Security scan results are archived and historically trackable",
            "evidence": "Scan result storage, historical trend reports, vulnerability tracking database",
            "category": "Security Scanning"
        },
        {
            "id": "3.14",
            "requirement": "Rollback procedures are tested and documented",
            "evidence": "Rollback runbooks, rollback test results, disaster recovery procedures",
            "category": "Deployment Controls"
        },
        {
            "id": "3.15",
            "requirement": "Deployment logs are sent to centralised logging/SIEM",
            "evidence": "SIEM integration, deployment audit logs, log retention policies",
            "category": "Deployment Controls"
        },
        {
            "id": "3.16",
            "requirement": "Pipeline secrets are managed securely (not hardcoded in configs)",
            "evidence": "Secrets management integration, vault usage, secret rotation logs",
            "category": "Pipeline Security"
        },
        {
            "id": "3.17",
            "requirement": "Build reproducibility is ensured (deterministic builds where possible)",
            "evidence": "Build reproducibility documentation, dependency pinning, lockfiles",
            "category": "Artifact Security"
        },
        {
            "id": "3.18",
            "requirement": "Failed security scans trigger notifications to security team",
            "evidence": "Alert configuration, notification logs, incident escalation procedures",
            "category": "Security Scanning"
        },
    ]


def get_domain4_requirements():
    """
    Domain 4: Security Testing Integration
    
    Testing is where we validate that our security controls actually work.
    Without testing, security is just wishful thinking.
    """
    return [
        {
            "id": "4.1",
            "requirement": "Security test cases are written for all critical functionality",
            "evidence": "Security test suite, test case repository, test coverage reports",
            "category": "Test Planning"
        },
        {
            "id": "4.2",
            "requirement": "Unit tests include security-focused test cases (input validation, etc.)",
            "evidence": "Unit test suites with security tests, code coverage reports",
            "category": "Unit Testing"
        },
        {
            "id": "4.3",
            "requirement": "Integration tests validate authentication and authorisation mechanisms",
            "evidence": "Integration test suites, auth test results, test scenarios",
            "category": "Integration Testing"
        },
        {
            "id": "4.4",
            "requirement": "DAST (Dynamic Application Security Testing) runs against staging environments",
            "evidence": "DAST scan reports (OWASP ZAP, Burp Suite, Acunetix), vulnerability findings",
            "category": "Dynamic Testing"
        },
        {
            "id": "4.5",
            "requirement": "API security testing is performed (fuzzing, schema validation, injection tests)",
            "evidence": "API test results, fuzzing reports, Postman/REST Assured test suites",
            "category": "API Testing"
        },
        {
            "id": "4.6",
            "requirement": "Penetration testing is conducted for high-risk applications annually",
            "evidence": "Pentest reports, findings tracking, remediation validation",
            "category": "Penetration Testing"
        },
        {
            "id": "4.7",
            "requirement": "Security regression tests prevent reintroduction of known vulnerabilities",
            "evidence": "Regression test suite, vulnerability fix validation tests",
            "category": "Regression Testing"
        },
        {
            "id": "4.8",
            "requirement": "Test environments mirror production security configurations",
            "evidence": "Environment parity documentation, configuration comparisons",
            "category": "Test Environment"
        },
        {
            "id": "4.9",
            "requirement": "Automated security tests run in CI/CD pipeline on every build",
            "evidence": "Pipeline test execution logs, automated test results",
            "category": "Test Automation"
        },
        {
            "id": "4.10",
            "requirement": "Security testing results are reviewed before release approval",
            "evidence": "Release gate checklists, security sign-off records, test result reviews",
            "category": "Test Planning"
        },
        {
            "id": "4.11",
            "requirement": "Performance testing includes assessment of security control overhead",
            "evidence": "Load test results with security controls enabled, performance baselines",
            "category": "Performance Testing"
        },
        {
            "id": "4.12",
            "requirement": "Bug bounty or responsible disclosure program provides external testing",
            "evidence": "Bug bounty platform reports (HackerOne, Bugcrowd), vulnerability submissions",
            "category": "External Testing"
        },
        {
            "id": "4.13",
            "requirement": "Security Champions participate in test planning and review",
            "evidence": "Test planning meeting notes, Security Champion involvement records",
            "category": "Test Planning"
        },
        {
            "id": "4.14",
            "requirement": "Failed security tests block deployment to production",
            "evidence": "Quality gate configuration, blocked deployment logs, override procedures",
            "category": "Test Automation"
        },
        {
            "id": "4.15",
            "requirement": "Security test coverage is measured and tracked over time",
            "evidence": "Coverage metrics, security test inventory, trend analysis",
            "category": "Test Planning"
        },
        {
            "id": "4.16",
            "requirement": "Authentication and session management are thoroughly tested",
            "evidence": "Auth test scenarios, session security tests, credential handling tests",
            "category": "Integration Testing"
        },
        {
            "id": "4.17",
            "requirement": "Input validation testing covers all user input vectors",
            "evidence": "Input validation test cases, injection testing results, fuzzing outputs",
            "category": "API Testing"
        },
        {
            "id": "4.18",
            "requirement": "Security test data is maintained separately from production data",
            "evidence": "Test data management procedures, synthetic data generation tools",
            "category": "Test Environment"
        },
    ]


def get_domain5_requirements():
    """
    Domain 5: Release & Change Management
    
    The final gate before production. This is where we ensure that all
    security checks have passed and changes are controlled.
    """
    return [
        {
            "id": "5.1",
            "requirement": "Security review is mandatory for all production releases",
            "evidence": "Release checklists with security sign-off, security review records",
            "category": "Release Process"
        },
        {
            "id": "5.2",
            "requirement": "Release notes document all security-relevant changes",
            "evidence": "Release notes, change logs with security tags, security update notifications",
            "category": "Release Documentation"
        },
        {
            "id": "5.3",
            "requirement": "Emergency hotfix process includes security assessment",
            "evidence": "Hotfix procedures, emergency change records, security review logs",
            "category": "Change Management"
        },
        {
            "id": "5.4",
            "requirement": "Rollback plans are prepared and tested for all production releases",
            "evidence": "Rollback procedures, rollback test results, disaster recovery plans",
            "category": "Release Process"
        },
        {
            "id": "5.5",
            "requirement": "Post-deployment verification includes security checks",
            "evidence": "Smoke test results, security validation logs, post-deployment checklists",
            "category": "Release Process"
        },
        {
            "id": "5.6",
            "requirement": "Security patches are prioritised and tracked with defined SLAs",
            "evidence": "Security patch SLA, vulnerability remediation tracker, patching metrics",
            "category": "Patch Management"
        },
        {
            "id": "5.7",
            "requirement": "Configuration changes follow formal change management process",
            "evidence": "Change requests, CAB approval records, configuration change logs",
            "category": "Change Management"
        },
        {
            "id": "5.8",
            "requirement": "Production access is logged, monitored, and reviewed",
            "evidence": "Production access logs, privileged access monitoring, access reviews",
            "category": "Access Control"
        },
        {
            "id": "5.9",
            "requirement": "Feature flags enable controlled rollout of security features",
            "evidence": "Feature flag configurations, gradual rollout plans, feature toggle logs",
            "category": "Release Process"
        },
        {
            "id": "5.10",
            "requirement": "Security metrics are collected and reviewed post-release",
            "evidence": "Post-release security reports, incident metrics, security KPI dashboards",
            "category": "Release Documentation"
        },
        {
            "id": "5.11",
            "requirement": "Communication plan exists for security-relevant releases",
            "evidence": "Communication templates, stakeholder notification logs, security bulletins",
            "category": "Release Documentation"
        },
        {
            "id": "5.12",
            "requirement": "Release frequency balances security needs with business velocity",
            "evidence": "Release calendar, security vs. feature delivery metrics, deployment frequency",
            "category": "Release Process"
        },
        {
            "id": "5.13",
            "requirement": "Post-incident releases are tracked and undergo enhanced scrutiny",
            "evidence": "Incident-driven release logs, security fix tracking, post-incident reviews",
            "category": "Patch Management"
        },
        {
            "id": "5.14",
            "requirement": "Compliance artifacts are updated with each relevant release",
            "evidence": "Compliance documentation updates, attestation records, audit trail",
            "category": "Release Documentation"
        },
        {
            "id": "5.15",
            "requirement": "Lessons learned from security issues feed into process improvements",
            "evidence": "Retrospective notes, security improvement backlog, process change records",
            "category": "Continuous Improvement"
        },
        {
            "id": "5.16",
            "requirement": "Release approvals require sign-off from security stakeholders",
            "evidence": "Approval workflow records, security sign-off documentation",
            "category": "Release Process"
        },
        {
            "id": "5.17",
            "requirement": "Version control and release tags are properly maintained",
            "evidence": "Git tags, release versioning records, artifact version tracking",
            "category": "Release Documentation"
        },
        {
            "id": "5.18",
            "requirement": "Security scan results are archived with release artifacts",
            "evidence": "Scan result archives, release documentation bundles, compliance packages",
            "category": "Release Documentation"
        },
    ]

# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET CREATION
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each assessment domain sheet sequentially.', '2. For each requirement, select Implementation Status from the dropdown menu.', '3. Provide Evidence Reference documenting WHERE the evidence can be found.', '4. Add Comments explaining implementation details, exceptions, or context.', '5. Review the Summary Dashboard to see overall compliance metrics.', '6. Document all gaps in the Gap Analysis sheet with remediation plans.', '7. Maintain the Evidence Register for comprehensive audit trail.', '8. Obtain required approvals on the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_domain_sheet(ws, domain_name, requirements, styles):
    """
    Create a standardized assessment domain sheet.
    
    Each domain sheet follows the same structure:
    - Header row with column names
    - Assessment items with ID, Requirement, Status dropdown, Evidence, Comments
    - Compliance indicator (auto-calculated based on status)
    
    Args:
        ws: Worksheet object
        domain_name: Display name for the domain
        requirements: List of requirement dictionaries
        styles: Style templates dictionary
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = f"SDLC Assessment: {domain_name}".upper()
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO 27001:2022 | Control A.8.28 | {domain_name}"
    ws["A2"].font = Font(italic=True, size=10, color="003366", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers
    headers = [
        "ID",
        "Requirement",
        "Implementation Status",
        "Evidence Reference",
        "Comments"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        apply_style(cell, styles["column_header"])
    
    ws.row_dimensions[3].height = 30

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40

    # Requirements data
    current_category = None
    row = 4
    
    for req in requirements:
        # Insert category header if new category
        if "category" in req and req["category"] != current_category:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"▶ {req['category']}"
            ws[f"A{row}"].font = Font(bold=True, size=11, color="4472C4")
            ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            current_category = req["category"]
        
        # ID column
        ws[f"A{row}"] = req["id"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Requirement column
        ws[f"B{row}"] = req["requirement"]
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Implementation Status (dropdown - yellow input cell)
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        validations["implementation_status"].add(ws[f"C{row}"])
        
        # Evidence Reference (yellow input cell)
        ws[f"D{row}"] = req.get("evidence", "")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"D{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[f"D{row}"].font = Font(size=9, italic=True, color="666666")
        
        # Comments (yellow input cell)
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"E{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row += 1

    # Pad to 50 FFFFCC data rows for QA compliance (DS-005)
    min_data_rows = 50
    data_rows_written = len(requirements)
    if data_rows_written < min_data_rows:
        thin = Side(style="thin")
        for pad_row in range(row, row + (min_data_rows - data_rows_written)):
            for col in ["A", "B", "C", "D", "E"]:
                cell = ws[f"{col}{pad_row}"]
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws[f"C{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"C{pad_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            validations["implementation_status"].add(ws[f"C{pad_row}"])
            ws[f"D{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"D{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws[f"E{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"E{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Hide unused columns (F onwards) to prevent empty column display
    for col_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col_letter].hidden = True

    # Freeze header rows
    ws.freeze_panes = "A4"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    return ws


def add_grey_info_rows(ws, items):
    """
    Add F2F2F2 grey informational rows after the last requirement row.

    These rows are visible reference information and are NOT counted by COUNTIF
    formulas (which target FFFFCC input cells only). They overwrite the first
    N FFFFCC padding rows written by create_domain_sheet().

    Strategy: find the last row where column B has content (last requirement),
    then the next row is the first FFFFCC padding row — overwrite with F2F2F2.
    """
    thin = Side(style="thin")
    # Find last row with content in column B (last requirement row)
    last_req_row = 2  # default to header row
    for row_idx in range(ws.max_row, 2, -1):
        cell_b = ws.cell(row=row_idx, column=2)
        if cell_b.value is not None and str(cell_b.value).strip() != "":
            last_req_row = row_idx
            break

    # Overwrite the first len(items) FFFFCC padding rows with F2F2F2
    for i, item_text in enumerate(items):
        r = last_req_row + 1 + i
        for col in range(1, 6):  # columns A-E
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.cell(row=r, column=2).value = item_text
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, color="003366", italic=True)
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        # Clear any FFFFCC fill or DV from col C (it was a padding input cell)
        ws.cell(row=r, column=3).fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )


# ============================================================================
# SECTION 6: SUMMARY DASHBOARD CREATION
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard matching A.8.33 gold standard with TABLE 1, 2, 3."""
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (italic, 003366 font, no fill)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 Annex A.8.28 — Secure Coding | SDLC Assessment"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4: TABLE 1 banner (003366, merged A4:G4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=4, column=col).border = border

    # Row 5: Column headers (D9D9D9 grey, black text)
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows (4-8) - one per assessment area
    area_configs = [
        ("Security Requirements Design", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Development Environment", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Build Deployment Pipeline", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Security Testing Integration", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Release Change Management", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
    ]

    for i, (area_name, status_col, status_values) in enumerate(area_configs):
        row = 6 + i

        # Column A: Area name
        ws.cell(row=row, column=1, value=area_name).border = border

        # Column B: Total Items (COUNTA of status column)
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTA('{area_name}'!{status_col}4:{status_col}100)"
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal="center")

        # Column C: Compliant (✅ Implemented)
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[0]}")'
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal="center")

        # Column D: Partial (⚠️ Partially Implemented)
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[1]}")'
        cell_d.border = border
        cell_d.alignment = Alignment(horizontal="center")

        # Column E: Non-Compliant (❌ Not Implemented)
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[2]}")'
        cell_e.border = border
        cell_e.alignment = Alignment(horizontal="center")

        # Column F: N/A count
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"N/A")'
        cell_f.border = border
        cell_f.alignment = Alignment(horizontal="center")

        # Column G: Compliance % (Compliant / (Total - N/A))
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f'=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))'
        cell_g.number_format = "0.0%"
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center")

    # TOTAL row (after data rows + 2 buffer rows)
    total_row = 6 + len(area_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1).border = border
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    # Compliance % formula
    cell = ws.cell(row=total_row, column=7)
    cell.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    cell.number_format = "0.0%"
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    # TABLE 2: Key Metrics
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{metrics_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=metrics_start, column=col).border = border

    # TABLE 2 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        hcell = ws.cell(row=metrics_start + 1, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center")

    # Metrics exploiting DVs from all 5 sheets
    metrics = [
        # Security Requirements & Design
        ("Total Requirements Tracked", "=COUNTA('Security Requirements Design'!B5:B100)"),
        ("Security Requirements Documented", '=COUNTIF(\'Security Requirements Design\'!C5:C100,"✅ Implemented")'),
        ("Threat Modeling Coverage", '=COUNTIF(\'Security Requirements Design\'!C5:C100,"✅ Implemented")/COUNTA(\'Security Requirements Design\'!B5:B100)'),
        # Development Environment
        ("Total Dev Environment Controls", "=COUNTA('Development Environment'!B5:B100)"),
        ("Workstations with Security Tools", '=COUNTIF(\'Development Environment\'!C5:C100,"✅ Implemented")'),
        ("Repository Security Controls", '=COUNTIF(\'Development Environment\'!C5:C100,"✅ Implemented")'),
        # Build & Deployment
        ("Total Pipeline Controls", "=COUNTA('Build Deployment Pipeline'!B5:B100)"),
        ("Security Scans Integrated", '=COUNTIF(\'Build Deployment Pipeline\'!C5:C100,"✅ Implemented")'),
        ("Deployment Gates Active", '=COUNTIF(\'Build Deployment Pipeline\'!C5:C100,"✅ Implemented")'),
        # Security Testing
        ("Total Testing Controls", "=COUNTA('Security Testing Integration'!B5:B100)"),
        ("SAST/DAST Integration", '=COUNTIF(\'Security Testing Integration\'!C5:C100,"✅ Implemented")'),
        ("Penetration Testing Coverage", '=COUNTIF(\'Security Testing Integration\'!C5:C100,"✅ Implemented")'),
        # Release & Change Management
        ("Total Release Controls", "=COUNTA('Release Change Management'!B5:B100)"),
        ("Approval Workflows Active", '=COUNTIF(\'Release Change Management\'!C5:C100,"✅ Implemented")'),
        ("Change Management Compliance", '=COUNTIF(\'Release Change Management\'!C5:C100,"✅ Implemented")/COUNTA(\'Release Change Management\'!B5:B100)'),
    ]

    r = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=r, column=1, value=metric).border = border
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10)
        cell_val = ws.cell(row=r, column=2, value=formula)
        cell_val.border = border
        cell_val.font = Font(name="Calibri", size=10, color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # 2 buffer rows after TABLE 2 metrics (white, bordered)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # TABLE 3: Critical Findings (red banner)
    crit_start = r + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    # TABLE 3 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Critical findings based on A.8.28 secure coding requirements
    findings = [
        ("Security Requirements Design", "Requirements without security acceptance criteria", '=COUNTIF(\'Security Requirements Design\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Development Environment", "Developer workstations without EDR/security tools", '=COUNTIF(\'Development Environment\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Build Deployment Pipeline", "Pipelines without SAST/SCA scanning", '=COUNTIF(\'Build Deployment Pipeline\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Security Testing Integration", "Production code without security testing", '=COUNTIF(\'Security Testing Integration\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Release Change Management", "Releases without security approval", '=COUNTIF(\'Release Change Management\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Security Requirements Design", "Missing threat modeling for high-risk apps", '=COUNTIF(\'Security Requirements Design\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Development Environment", "Repositories without MFA enforcement", '=COUNTIF(\'Development Environment\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Build Deployment Pipeline", "Security gates not enforced", '=COUNTIF(\'Build Deployment Pipeline\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Security Testing Integration", "Incomplete DAST coverage", '=COUNTIF(\'Security Testing Integration\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Release Change Management", "Emergency changes without post-review", '=COUNTIF(\'Release Change Management\'!C5:C100,"⚠️ Partially Implemented")', "Medium", "Plan"),
    ]

    r3 = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        ws.cell(row=r3, column=1, value=cat)
        ws.cell(row=r3, column=2, value=finding)
        count_cell = ws.cell(row=r3, column=3, value=formula)
        count_cell.alignment = Alignment(horizontal="center")
        count_cell.font = Font(name="Calibri", size=10, color="000000")
        ws.cell(row=r3, column=4, value=severity)
        ws.cell(row=r3, column=5, value=action)
        r3 += 1

    # 2 buffer rows after TABLE 3 (FFFFCC)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        r3 += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"  # Freeze below subtitle + banner + headers

    return ws


# ============================================================================
# SECTION 7: EVIDENCE REGISTER CREATION
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        ("Evidence ID", 15),
        ("Assessment Area", 25),
        ("Evidence Type", 22),
        ("Description", 40),
        ("Location/Path", 45),
        ("Date Collected", 16),
        ("Collected By", 20),
        ("Verification Status", 22),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border_thin
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Data rows EV-001 to EV-100
    dv_etype = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=False
    )
    dv_vstatus = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False
    )
    ws.add_data_validation(dv_etype)
    ws.add_data_validation(dv_vstatus)

    # MAX-001/MAX-002 fix: 1 grey sample + 100 FFFFCC empty rows
    # Sample row (row 5) with F2F2F2 grey background (like A.8.33 reference)
    sample_data = [
        "EV-001",
        "Test Data Inventory",
        "Configuration file",
        "Data masking configuration for UAT environment",
        "\\fileserver\\isms\\evidence\\test-data\\uat-masking-config.json",
        "2024-03-01",
        "Sarah Mitchell",
        "Verified"
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=5, column=col_num, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = border_thin

    # 100 empty rows (rows 6-105) - FFFFCC yellow
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
        dv_etype.add(ws.cell(row=r, column=3))
        dv_vstatus.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"
    return ws

# ============================================================================
# SECTION 8: GAP ANALYSIS SHEET CREATION
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """
    Create Gap Analysis sheet for tracking remediation efforts.
    
    Identifying gaps is easy. Fixing them requires planning, ownership, and tracking.
    This sheet turns assessment findings into actionable remediation plans.
    """
    validations = create_base_validations(ws)
    
    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "GAP ANALYSIS & REMEDIATION PLAN"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all 'Not Implemented' and 'Partially Implemented' requirements. Assign owners, set target dates, and track remediation progress."
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column headers
    headers = [
        "Gap ID",
        "Domain",
        "Requirement ID",
        "Requirement Description",
        "Current State",
        "Target State",
        "Priority",
        "Owner",
        "Target Date",
        "Status",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Column widths
    widths = {
        "A": 10,
        "B": 22,
        "C": 12,
        "D": 35,
        "E": 25,
        "F": 25,
        "G": 12,
        "H": 20,
        "I": 12,
        "J": 15,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Add dropdowns
    validations["priority"].add("G4:G200")
    validations["gap_status"].add("J4:J200")

    # MAX-003 fix: 1 sample row (F2F2F2 grey) + 50 empty rows (FFFFCC yellow) per Option B standard
    thin = Side(style="thin")

    # Sample row (row 4) - F2F2F2 grey with realistic example data
    sample_data = [
        "GAP-001",
        "Development Environment",
        "2.4",
        "Pre-commit hooks prevent secret commits",
        "No pre-commit hooks configured",
        "Git-secrets installed on all dev workstations",
        "High",
        "DevOps Lead",
        "31.01.2025",
        "In Progress"
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col_num, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 50 empty rows (rows 5-54) - FFFFCC yellow
    for gap_row in range(5, 55):
        for col in range(1, 11):
            cell = ws.cell(row=gap_row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 55

    # Add conditional formatting guidance below examples
    row += 1
    ws[f"A{row}"] = "Priority Guidance:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"A{row}:J{row}")
    
    row += 1
    priority_guide = [
        "Critical: Exploitable vulnerability or compliance blocker - fix immediately",
        "High: Significant security risk or major compliance gap - fix within 30 days",
        "Medium: Moderate security risk or minor compliance gap - fix within 90 days",
        "Low: Low security risk or improvement opportunity - fix when resources available",
    ]
    
    for guide in priority_guide:
        ws[f"A{row}"] = guide
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:J{row}")
        
        # Color code based on priority
        if "Critical" in guide:
            ws[f"A{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif "High" in guide:
            ws[f"A{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif "Medium" in guide:
            ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif "Low" in guide:
            ws[f"A{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        row += 1

    # Hide unused columns (K onwards) to prevent empty column display in Gap Analysis
    for col_letter in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    return ws


# ============================================================================
# SECTION 9: APPROVAL SIGN-OFF SHEET CREATION
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=1, column=col).border = border_thin
    ws.row_dimensions[1].height = 35

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=2, column=col).border = border_thin

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=row, column=col).border = border_thin

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.28.1 - SDLC Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        # Apply borders to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = border_thin
            if value == "":
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Status dropdown on Assessment Status
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Apply borders to banner (A:E)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_thin
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            # Apply borders and fill to all cells in merged range (B:E)
            for col in range(2, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=row, column=col).border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all cells in merged range (B:E)
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border_thin

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to banner (A:E)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border_thin

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

    return ws



# ============================================================================
# SECTION 10: MAIN EXECUTION
# ============================================================================

def main():
    """
    Main execution function - generates complete SDLC Assessment workbook.
    
    As Feynman taught us: "What I cannot create, I do not understand."
    This script creates a complete, functional assessment workbook that
    actually helps organisations understand their SDLC security posture.
    """
    logger.info("=" * 80)
    logger.info(" " * 20 + "ISMS Control 8.28.1 - SDLC Assessment Generator")
    logger.info("=" * 80)
    logger.info("")
    logger.info("Generating comprehensive SDLC security assessment workbook...")
    logger.info("")
    
    try:
        # Create workbook structure
        logger.info("[D] [1/10] Creating workbook structure...")
        wb = create_workbook()
        styles = _STYLES
        logger.info("     \u2705 Workbook initialized with 10 sheets")

        # Create Instructions sheet
        logger.info("[F] [2/10] Creating Instructions sheet...")
        ws_instructions = wb["Instructions & Legend"]
        ws_instructions.sheet_view.showGridLines = False
        create_instructions_sheet(ws_instructions)
        logger.info("     \u2705 Instructions sheet complete (guidance, legend, examples)")

        # Create assessment domain sheets
        logger.info("[CHART] [3/10] Creating Domain 1: Security Requirements & Design...")
        ws_domain1 = wb["Security Requirements Design"]
        ws_domain1.sheet_view.showGridLines = False
        create_domain_sheet(
            ws_domain1,
            "Security Requirements & Design",
            get_domain1_requirements(),
            styles
        )
        logger.info("     \u2705 18 requirements (Requirements Definition, Design Review)")

        logger.info("[CHART] [4/10] Creating Domain 2: Development Environment & Tooling...")
        ws_domain2 = wb["Development Environment"]
        ws_domain2.sheet_view.showGridLines = False
        create_domain_sheet(
            ws_domain2,
            "Development Environment & Tooling",
            get_domain2_requirements(),
            styles
        )
        logger.info("     \u2705 18 requirements (Workstation Security, Repository Security, Tools)")

        logger.info("[CHART] [5/10] Creating Domain 3: Build & Deployment Pipeline...")
        ws_domain3 = wb["Build Deployment Pipeline"]
        ws_domain3.sheet_view.showGridLines = False
        create_domain_sheet(
            ws_domain3,
            "Build & Deployment Pipeline",
            get_domain3_requirements(),
            styles
        )
        # ISO 27002:2022 Phase 2 content: add supply chain security reference rows (F2F2F2)
        add_grey_info_rows(ws_domain3, [
            "SBOM (Software Bill of Materials) generated and archived for each release",
            "Dependency pinning enforced — lockfiles committed, no floating version ranges "
            "in production dependencies",
            "Supply chain attack detection — build artefact integrity verified "
            "(checksums, signed artefacts)",
        ])
        logger.info("     \u2713 18 requirements + 3 reference rows (Pipeline Security, Security Scanning, Deployment)")

        logger.info("[CHART] [6/10] Creating Domain 4: Security Testing Integration...")
        ws_domain4 = wb["Security Testing Integration"]
        ws_domain4.sheet_view.showGridLines = False
        create_domain_sheet(
            ws_domain4,
            "Security Testing Integration",
            get_domain4_requirements(),
            styles
        )
        logger.info("     \u2705 18 requirements (Unit/Integration/API/DAST/Pentest)")

        logger.info("[CHART] [7/10] Creating Domain 5: Release & Change Management...")
        ws_domain5 = wb["Release Change Management"]
        ws_domain5.sheet_view.showGridLines = False
        create_domain_sheet(
            ws_domain5,
            "Release & Change Management",
            get_domain5_requirements(),
            styles
        )
        logger.info("     \u2705 18 requirements (Release Process, Patch Management, Approval)")

        logger.info("[TREND] [8/10] Creating Summary Dashboard...")
        ws_summary = wb["Summary Dashboard"]
        ws_summary.sheet_view.showGridLines = False
        create_summary_dashboard_sheet(ws_summary, styles)
        logger.info("     \u2705 Executive summary with compliance metrics and traffic lights")

        logger.info("[SEARCH] [9/10] Creating Evidence Register...")
        ws_evidence = wb["Evidence Register"]
        ws_evidence.sheet_view.showGridLines = False
        create_evidence_register(ws_evidence, styles)
        logger.info("     \u2705 Evidence tracking with examples and templates")

        logger.info("\u26A0\uFE0F  [10/10] Creating Gap Analysis & Approval sheets...")
        ws_gap = wb["Gap Analysis"]
        ws_gap.sheet_view.showGridLines = False
        create_gap_analysis_sheet(ws_gap, styles)
        
        ws_approval = wb["Approval Sign-Off"]
        ws_approval.sheet_view.showGridLines = False
        create_approval_sheet(ws_approval, styles)
        logger.info("     \u2705 Gap remediation tracking and formal approval workflow")

        # Save workbook
        logger.info("")
        logger.info("[S] Saving workbook...")
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.8.28.1_SDLC_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        # Success summary
        logger.info("")
        logger.info("=" * 80)
        logger.info("\u2705 SUCCESS: SDLC Assessment workbook generated successfully!")
        logger.info("=" * 80)
        logger.info("")
        logger.info(f"[D] File: {filename}")
        logger.info(f"[CHART] Sheets: {len(wb.sheetnames)}")
        logger.info(f"\u25A1 Total Requirements: 90 (18 per domain × 5 domains)")
        logger.info("")
        logger.info("[PIN] Assessment Domains:")
        logger.info("   1. Security Requirements & Design        (18 requirements)")
        logger.info("   2. Development Environment & Tooling     (18 requirements)")
        logger.info("   3. Build & Deployment Pipeline           (18 requirements)")
        logger.info("   4. Security Testing Integration          (18 requirements)")
        logger.info("   5. Release & Change Management           (18 requirements)")
        logger.info("")
        logger.info("[PIN] Supporting Sheets:")
        logger.info("   \u2022 Instructions (comprehensive guidance)")
        logger.info("   \u2022 Summary Dashboard (executive overview)")
        logger.info("   \u2022 Evidence Register (audit trail)")
        logger.info("   \u2022 Gap Analysis (remediation tracking)")
        logger.info("   \u2022 Approval Sign-Off (formal approval)")
        logger.info("")
        logger.info("[TARGET] Next Steps:")
        logger.info("   1. Open the workbook and review the Instructions sheet")
        logger.info("   2. Complete each assessment domain systematically")
        logger.info("   3. Provide evidence references for all 'Implemented' items")
        logger.info("   4. Document gaps and create remediation plans")
        logger.info("   5. Review Summary Dashboard for compliance metrics")
        logger.info("   6. Obtain stakeholder approvals")
        logger.info("")
        logger.info("[TIP] Remember: Security is not cargo cult. Focus on what is ACTUALLY")
        logger.info("   implemented and working, not just what exists on paper.")
        logger.info("")
        logger.info("=" * 80)
        
        return 0

    except Exception as e:
        logger.info("")
        logger.info("=" * 80)
        logger.error("\u274C ERROR: Failed to generate workbook")
        logger.info("=" * 80)
        logger.error(f"Error Details: {str(e)}")
        logger.info("")
        import traceback
        traceback.print_exc()
        logger.info("")
        logger.error("Please report this error to the ISMS implementation team.")
        logger.info("=" * 80)
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
