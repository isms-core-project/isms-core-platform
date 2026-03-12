#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.28.4 - Third-Party & Open Source Software Assessment
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Assessment Domain 4 of 4: Third-Party Code and Open Source Component Security

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific dependency management practices, vendor evaluation
criteria, and open source governance policies.

Key customisation areas:
1. Dependency management tools and processes (match your package managers)
2. Open source approval criteria and risk thresholds (adapt to risk tolerance)
3. Vendor security assessment requirements (specific to procurement policies)
4. License compliance requirements (based on legal/IP policies)
5. Vulnerability remediation SLAs (aligned with risk management framework)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMISE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
third-party code security and open source component risk management across all
applications and development projects.

**Purpose:**
Enables systematic assessment of third-party and open source software security
against ISO 27001:2022 Control A.8.28 requirements, supporting evidence-based
validation of supply chain security and dependency risk management practices.

**Assessment Scope:**
- Software Composition Analysis (SCA) tool implementation and coverage
- Open source component inventory and version tracking
- Vulnerability scanning for dependencies and transitive dependencies
- License compliance tracking and approval workflows
- Third-party vendor security assessment processes
- Commercial/COTS software security evaluation
- Dependency update policies and patch management
- End-of-life (EOL) and unmaintained component identification
- Software Bill of Materials (SBOM) generation and maintenance
- Supply chain attack risk mitigation (dependency confusion, typosquatting)
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and OSS security standards
2. SCA Tool Coverage - Software Composition Analysis implementation assessment
3. Component Inventory - Open source and third-party component catalog
4. Vulnerability Management - Dependency vulnerability tracking and remediation
5. License Compliance - OSS license tracking and approval workflow
6. Vendor Assessment - Third-party vendor security evaluation process
7. Update Policy - Dependency update cadence and patch management
8. EOL Components - End-of-life and unmaintained dependency identification
9. SBOM Management - Software Bill of Materials generation and accuracy
10. Supply Chain Risk - Dependency confusion and supply chain attack mitigation
11. Gap Analysis - Inadequate controls or high-risk dependencies and remediation
12. Evidence Register - Audit evidence tracking and documentation
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dependency type and risk level dropdown lists
- Conditional formatting for vulnerability severity and remediation status
- Automated gap identification for high-risk or EOL dependencies
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with SCA tool outputs and vulnerability databases

**Integration:**
This assessment feeds into the A.8.28.5 Compliance Dashboard, which
consolidates data from all four assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a828_4_third_party_oss.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a828_4_third_party_oss.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a828_4_third_party_oss.py --date 20250124

Output:
    File: ISMS_A_8_28_4_Third_Party_OSS_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise OSS governance criteria to match your risk profile
    2. Inventory all applications and their dependency manifests
    3. Complete assessments for each application or microservice
    4. Validate SCA tool coverage and vulnerability detection accuracy
    5. Review EOL component list and plan migration/updates
    6. Conduct gap analysis for high-risk dependencies or missing controls
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (SCA reports, SBOMs, vendor assessments)
    9. Obtain stakeholder approvals (AppSec, Legal/Compliance, Procurement)
    10. Feed results into A.8.28.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    4 of 4 (Third-Party & OSS)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Implementation Guide
    - ISMS-IMP-A.8.28.1: SDLC Integration Assessment (Domain 1)
    - ISMS-IMP-A.8.28.2: Standards & Tools Assessment (Domain 2)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Assessment (Domain 3)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a828_1_sdlc_assessment.py
    - generate_a828_2_standards_tools.py
    - generate_a828_3_code_review_testing.py
    - generate_a828_5_compliance_dashboard.py
    - normalize_assessment_files_a828.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.28.4 specification
    - Supports comprehensive third-party and OSS security evaluation
    - Integrated with A.8.28.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Supply Chain Security Reality:**
Third-party and open source dependencies are the #1 attack vector in modern
software. SolarWinds, Log4Shell, event-stream npm package - all supply chain
attacks. This assessment domain is not optional or "nice to have" - it's
critical infrastructure security.

**Open Source ≠ Insecure:**
Don't treat all OSS as risky. Well-maintained OSS projects (Linux, OpenSSL,
React, etc.) often have better security than commercial alternatives due to
transparency and community review. Focus on maintenance activity, not just
presence of CVEs.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of SCA tool coverage, SBOM accuracy, and
vulnerability remediation for high/critical findings.

**Data Protection:**
Assessment workbooks contain sensitive supply chain information including:
- Complete dependency manifests and software architecture
- Vulnerability details and exploitability assessments
- Vendor security evaluation results and risk ratings
- License compliance issues and legal considerations

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Weekly: Check for new high/critical vulnerabilities in dependencies
- Monthly: Review EOL component list and update dependencies
- Quarterly: Complete SCA tool coverage and accuracy validation
- Annually: Full reassessment of all applications and dependency policies
- Ad-hoc: When major supply chain incidents occur (Log4Shell-scale events)

**Quality Assurance:**
Have application security SMEs, software architects, and legal/compliance
teams validate assessments before using results for compliance reporting or
remediation decisions. Legal review is essential for license compliance.

**Regulatory Alignment:**
Ensure third-party software security aligns with applicable requirements:
- Payment processing: PCI DSS v4.0.1 third-party service provider requirements
- Healthcare: HIPAA Business Associate Agreements for software vendors
- Finance: Regional banking third-party risk management requirements
- Government: Software supply chain security (SBOM requirements, EO 14028)

Customise assessment criteria to include regulatory-specific requirements.

**Integration with Dependency Tools:**
Where possible, integrate assessment data with:
- SCA tools (Snyk, WhiteSource, Black Duck, Dependabot, etc.)
- Package managers (npm, Maven, pip, NuGet, Gradle, etc.)
- Vulnerability databases (NVD, GitHub Advisory, OSV, etc.)
- SBOM generation tools (Syft, CycloneDX, SPDX tools)
- License compliance platforms (FOSSA, FOSSology, etc.)

Automated tool integration reduces manual effort and improves accuracy.

**Don't Fool Yourself - Feynman Principle:**
Just because you have an SCA tool doesn't mean you're secure. Key questions:
- Do you actually know all your dependencies? (Including transitive ones?)
- When a critical vulnerability is announced, can you identify affected apps
  within hours, not weeks?
- Do you have a process to update dependencies, or just detect vulnerabilities?
- Are developers empowered to update dependencies, or blocked by bureaucracy?
- Can you generate an accurate SBOM for any application on demand?

If you can't answer "yes" to these, you have dependency visibility problems,
not just dependency security problems. Fix visibility first.

**Dependency Update Philosophy:**
Staying current with dependencies is cheaper than emergency patching during
active exploitation. Regular dependency updates (monthly/quarterly) prevent
technical debt accumulation. Zero-day vulnerabilities in outdated dependencies
are especially dangerous because exploits are often public.

Automate dependency updates where possible (Dependabot, Renovate, etc.) and
treat dependency updates as normal maintenance, not exceptional events.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.28.4"
WORKBOOK_NAME = "Third-Party Code and Open Source Component Security"
CONTROL_ID = "A.8.28"
CONTROL_NAME = "Secure Coding"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")



import sys

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK INITIALIZATION AND STYLE DEFINITIONS
# ============================================================================

def create_workbook():
    """Initialize workbook with proper metadata."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Set workbook properties
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    return wb


def get_style_definitions():
    """
    Define standard style elements for consistent formatting.
    
    Colour Palette (Organisational Standard):
        - Headers: Dark blue (#003366) with white text
        - Subheaders: Medium blue (#4472C4) with white text
        - Section headers: Green (#70AD47)
        - Column headers: Light gray (#D9D9D9)
        - Input cells: Pale yellow (#FFFFCC)
        - Status colors: Green (#C6EFCE), Yellow (#FFEB9C), Red (#FFC7CE), Gray (#E7E6E6)
        - Priority colors: Dark red (#C00000), Light red (#FF6666), Yellow (#FFEB9C), Light green (#C6EFCE)
    """
    styles = {
        # Header styles
        'header_font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'header_fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        
        'subheader_font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
        'subheader_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        
        'section_font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'section_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        
        # Column header styles
        'column_header_font': Font(name='Calibri', size=10, bold=True),
        'column_header_fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        
        # Content styles
        'normal_font': Font(name='Calibri', size=10),
        'bold_font': Font(name='Calibri', size=10, bold=True),
        'italic_font': Font(name='Calibri', size=10, italic=True),
        
        # Input cell style
        'input_fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        
        # Status color fills
        'status_green': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'status_yellow': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'status_red': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'status_gray': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        
        # Priority color fills
        'priority_critical_fill': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
        'priority_critical_font': Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        'priority_high_fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'priority_medium_fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'priority_low_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        
        # Alignment styles
        'align_left': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'align_center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'align_right': Alignment(horizontal='right', vertical='top', wrap_text=True),
        
        # Border styles
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'thick_border': Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        ),
    }
    
    return styles


def get_column_widths():
    """
    Define standard column widths for assessment sheets.

    Returns:
        dict: Column letter -> width mapping
    """
    return {
        'A': 8,   # ID column
        'B': 55,  # Requirement description
        'C': 22,  # Implementation Status dropdown
        'D': 35,  # Evidence Reference
        'E': 40,  # Comments/Notes
    }


# ============================================================================
# SECTION 2: DATA VALIDATION DEFINITIONS
# ============================================================================

def create_data_validations():
    """
    Create reusable data validation objects (dropdowns) for the workbook.
    
    Returns:
        dict: Validation name -> DataValidation object mapping
    """
    validations = {}
    
    # Implementation Status dropdown (most common)
    validations['implementation_status'] = DataValidation(
        type="list",
        formula1='"✅ Implemented,⚠️ Partially Implemented,❌ Not Implemented,N/A"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select a valid implementation status from the dropdown."
    )
    validations['implementation_status'].prompt = "Select implementation status"
    validations['implementation_status'].promptTitle = "Implementation Status"
    
    # Yes/No dropdown
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    validations['yes_no'].prompt = "Select Yes or No"
    validations['yes_no'].promptTitle = "Yes/No"
    
    # Yes/No/N/A dropdown
    validations['yes_no_na'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    validations['yes_no_na'].prompt = "Select Yes, No, or N/A"
    validations['yes_no_na'].promptTitle = "Yes/No/N/A"
    
    # Priority dropdown
    validations['priority'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    validations['priority'].prompt = "Select priority level"
    validations['priority'].promptTitle = "Priority"
    
    # Gap Status dropdown (for Gap Analysis sheet)
    validations['gap_status'] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Closed,Deferred"',
        allow_blank=False
    )
    validations['gap_status'].prompt = "Select gap status"
    validations['gap_status'].promptTitle = "Gap Status"
    
    # Evidence Type dropdown
    validations['evidence_type'] = DataValidation(
        type="list",
        formula1='"Document,Screenshot,Report,Configuration,URL,Recording,Database Query,Log File,Other"',
        allow_blank=False
    )
    validations['evidence_type'].prompt = "Select evidence type"
    validations['evidence_type'].promptTitle = "Evidence Type"
    
    # Verification Status dropdown (for Evidence Register)
    validations['verification_status'] = DataValidation(
        type="list",
        formula1='"Pending,Verified,Rejected"',
        allow_blank=False
    )
    validations['verification_status'].prompt = "Select verification status"
    validations['verification_status'].promptTitle = "Verification Status"
    
    # Approval Decision dropdown (for Approval Sign-Off)
    validations['approval_decision'] = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
        allow_blank=False
    )
    validations['approval_decision'].prompt = "Select approval decision"
    validations['approval_decision'].promptTitle = "Approval Decision"
    
    # Domain dropdown (for Gap Analysis)
    validations['domain'] = DataValidation(
        type="list",
        formula1='"Vendor Security Assessment,OSS Management,Dependency Vulnerability Mgmt,Third-Party Code & Integration,License Compliance & Legal Risk"',
        allow_blank=False
    )
    validations['domain'].prompt = "Select assessment domain"
    validations['domain'].promptTitle = "Domain"
    
    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
def apply_validation_to_range(ws, validation, start_row, end_row, column):
    """
    Apply a data validation to a range of cells.
    
    Args:
        ws: Worksheet object
        validation: DataValidation object
        start_row: Starting row number
        end_row: Ending row number
        column: Column letter (e.g., 'C')
    """
    ws.add_data_validation(validation)
    for row in range(start_row, end_row + 1):
        cell_ref = f"{column}{row}"
        validation.add(cell_ref)


def apply_standard_formatting(ws, start_row, end_row, styles, column_widths):
    """
    Apply standard formatting to assessment domain rows.
    
    Args:
        ws: Worksheet object
        start_row: First data row
        end_row: Last data row
        styles: Style definitions dictionary
        column_widths: Column width definitions
    """
    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply borders and alignment to all cells
    for row in range(start_row, end_row + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.border = styles['thin_border']
            cell.font = styles['normal_font']
            
            # Column-specific alignment
            if col == 'A':
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif col == 'F':
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = styles['align_left']
    
    # Highlight input columns (C, D, E) with pale yellow
    for row in range(start_row, end_row + 1):
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].fill = styles['input_fill']


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def set_header_row(ws, row_num, headers, styles):
    """
    Create a formatted header row.
    
    Args:
        ws: Worksheet object
        row_num: Row number for headers
        headers: List of header texts
        styles: Style definitions dictionary
    """
    for idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=idx)
        cell.value = header_text
        cell.font = styles['column_header_font']
        cell.fill = styles['column_header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']


def freeze_header_rows(ws, row_num):
    """
    Freeze rows above the specified row number.
    
    Args:
        ws: Worksheet object
        row_num: Row number to freeze above (e.g., 4 freezes rows 1-3)
    """
    ws.freeze_panes = f'A{row_num}'


def add_title_section(ws, row_num, title, styles):
    """
    Add a formatted title section spanning multiple columns.

    Args:
        ws: Worksheet object
        row_num: Row number for title
        title: Title text
        styles: Style definitions dictionary

    Returns:
        int: Next available row number
    """
    ws.merge_cells(f'A{row_num}:E{row_num}')
    cell = ws[f'A{row_num}']
    cell.value = title
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = styles['align_center']
    cell.border = styles['thick_border']

    return row_num + 1


def add_metadata_row(ws, row_num, label, value, styles):
    """
    Add a metadata row (e.g., "Date:                 [Date to be set]").
    
    Args:
        ws: Worksheet object
        row_num: Row number
        label: Label text (e.g., "Assessment Date:")
        value: Value text
        styles: Style definitions dictionary
    
    Returns:
        int: Next available row number
    """
    ws[f'A{row_num}'] = label
    ws[f'A{row_num}'].font = styles['bold_font']
    ws[f'A{row_num}'].alignment = styles['align_left']
    
    ws.merge_cells(f'B{row_num}:F{row_num}')
    ws[f'B{row_num}'] = value
    ws[f'B{row_num}'].font = styles['normal_font']
    ws[f'B{row_num}'].alignment = styles['align_left']
    
    return row_num + 1

# ============================================================================
# SECTION 3: ASSESSMENT DOMAIN DATA DEFINITIONS
# ============================================================================

def get_vendor_security_requirements():
    """
    Domain 1: Vendor Security Assessment (18 requirements)

    Focus: Vendor due diligence, contracts, access controls, risk management

    Returns:
        list: List of dicts with keys: id, requirement, evidence, category
    """
    requirements = [
        {
            "id": "V-01",
            "requirement": "Vendor security assessment policy exists and is approved",
            "evidence": "Policy document, approval signatures, policy publication records",
            "category": "Vendor Policy",
        },
        {
            "id": "V-02",
            "requirement": "Security questionnaire sent to all vendors before onboarding",
            "evidence": "Questionnaire templates, completed questionnaires, vendor onboarding records",
            "category": "Vendor Assessment",
        },
        {
            "id": "V-03",
            "requirement": "Vendor risk classification framework implemented (High/Medium/Low tiers)",
            "evidence": "Risk classification matrix, vendor tier assignments, risk scoring methodology",
            "category": "Vendor Assessment",
        },
        {
            "id": "V-04",
            "requirement": "Contract security requirements defined for all vendors",
            "evidence": "Contract templates with security clauses, executed vendor contracts",
            "category": "Vendor Contracting",
        },
        {
            "id": "V-05",
            "requirement": "Data protection requirements specified in vendor contracts (DPAs, data handling)",
            "evidence": "Contract sections on data handling, Data Processing Agreements (DPAs), privacy clauses",
            "category": "Vendor Contracting",
        },
        {
            "id": "V-06",
            "requirement": "Incident response coordination requirements included in vendor contracts",
            "evidence": "Incident notification SLAs, contact procedures, escalation paths in contracts",
            "category": "Vendor Contracting",
        },
        {
            "id": "V-07",
            "requirement": "Vendor access controls documented and enforced (least privilege)",
            "evidence": "Access logs, permission matrices, SSO configurations, access request records",
            "category": "Vendor Access Control",
        },
        {
            "id": "V-08",
            "requirement": "Vendor accounts follow least privilege principle (minimal necessary permissions)",
            "evidence": "Access reviews, role definitions, permission audits, access justifications",
            "category": "Vendor Access Control",
        },
        {
            "id": "V-09",
            "requirement": "Vendor access reviewed regularly (quarterly for high-risk vendors)",
            "evidence": "Access review reports, approval records, review meeting notes, access certification logs",
            "category": "Vendor Access Control",
        },
        {
            "id": "V-10",
            "requirement": "High-risk vendors undergo annual comprehensive security reviews",
            "evidence": "Review meeting notes, security assessment reports, vendor scorecards, review schedules",
            "category": "Vendor Monitoring",
        },
        {
            "id": "V-11",
            "requirement": "Vendor security incidents tracked and reviewed (lessons learned)",
            "evidence": "Incident logs, post-mortem reports, lessons learned documents, remediation plans",
            "category": "Vendor Monitoring",
        },
        {
            "id": "V-12",
            "requirement": "Vendor security posture monitored (security ratings, breach notifications)",
            "evidence": "Security rating subscriptions (BitSight, SecurityScorecard), monitoring reports, breach alerts",
            "category": "Vendor Monitoring",
        },
        {
            "id": "V-13",
            "requirement": "Vendor SOC 2 / ISO 27001 compliance verified where applicable",
            "evidence": "SOC 2 Type II reports, ISO 27001 certificates, attestation letters, compliance verification",
            "category": "Vendor Compliance",
        },
        {
            "id": "V-14",
            "requirement": "Vendor data retention and deletion requirements specified in contracts",
            "evidence": "Contract data handling clauses, data deletion procedures, retention schedules",
            "category": "Vendor Compliance",
        },
        {
            "id": "V-15",
            "requirement": "Vendor security questionnaire responses validated (not blindly accepted)",
            "evidence": "Validation notes, follow-up correspondence, independent verification evidence",
            "category": "Vendor Assessment",
        },
        {
            "id": "V-16",
            "requirement": "Vendor offboarding process includes security checklist (access revocation, data return)",
            "evidence": "Offboarding checklists, access revocation logs, data deletion certificates",
            "category": "Vendor Offboarding",
        },
        {
            "id": "V-17",
            "requirement": "Critical vendors have business continuity plans reviewed and tested",
            "evidence": "BCP documents, disaster recovery plans, BCP test results, failover procedures",
            "category": "Business Continuity",
        },
        {
            "id": "V-18",
            "requirement": "Vendor security metrics tracked (% compliant vendors, review completion rates)",
            "evidence": "Vendor compliance dashboard, metrics reports, KPI tracking, trend analysis",
            "category": "Vendor Metrics",
        },
    ]

    return requirements


def get_oss_management_requirements():
    """
    Domain 2: Open Source Software Management (18 requirements)

    Focus: OSS approval, inventory (SBOM), license tracking, abandoned dependencies

    Returns:
        list: List of dicts with keys: id, requirement, evidence, category
    """
    requirements = [
        {
            "id": "O-01",
            "requirement": "OSS usage policy exists and is approved (defines acceptable OSS practices)",
            "evidence": "OSS policy document, approval signatures, policy distribution records",
            "category": "OSS Policy",
        },
        {
            "id": "O-02",
            "requirement": "OSS approval workflow defined (request → review → approve process)",
            "evidence": "Workflow documentation, JIRA/ServiceNow workflow configs, process diagrams",
            "category": "OSS Policy",
        },
        {
            "id": "O-03",
            "requirement": "OSS approval process enforced before production use (not bypassed)",
            "evidence": "Approval tickets, process audit logs, enforcement metrics, compliance reports",
            "category": "OSS Policy",
        },
        {
            "id": "O-04",
            "requirement": "OSS inventory maintained (SBOM) for all applications and services",
            "evidence": "SBOM exports (CycloneDX, SPDX format), inventory databases, dependency manifests",
            "category": "OSS Inventory",
        },
        {
            "id": "O-05",
            "requirement": "SBOM generated automatically in CI/CD pipeline (not manual)",
            "evidence": "CI/CD pipeline configs, build logs showing SBOM generation, automation scripts",
            "category": "OSS Inventory",
        },
        {
            "id": "O-06",
            "requirement": "OSS components tracked with specific version numbers (not version ranges)",
            "evidence": "Dependency manifests (package.json, requirements.txt, pom.xml), lockfiles",
            "category": "OSS Inventory",
        },
        {
            "id": "O-07",
            "requirement": "Transitive dependencies identified and tracked (dependencies of dependencies)",
            "evidence": "Dependency tree reports, SCA tool outputs showing full dependency graph",
            "category": "OSS Inventory",
        },
        {
            "id": "O-08",
            "requirement": "OSS license information captured in inventory for all dependencies",
            "evidence": "License scan reports, SBOM license fields, license identification logs",
            "category": "Dependency Health",
        },
        {
            "id": "O-09",
            "requirement": "Abandoned/unmaintained OSS detected and tracked (staleness monitoring)",
            "evidence": "Dependency staleness reports, EOL (end-of-life) tracking, maintainer activity checks",
            "category": "Dependency Health",
        },
        {
            "id": "O-10",
            "requirement": "OSS update policy defines acceptable staleness thresholds (e.g., <6 months since last update)",
            "evidence": "Policy document with staleness thresholds, update frequency requirements",
            "category": "Dependency Health",
        },
        {
            "id": "O-11",
            "requirement": "OSS components reviewed for security before approval (not rubber-stamped)",
            "evidence": "Security review checklists, approval criteria, security assessment notes",
            "category": "OSS Approval",
        },
        {
            "id": "O-12",
            "requirement": "Restricted licenses flagged during approval process (GPL, AGPL, SSPL)",
            "evidence": "License allowlist/blocklist, flagging rules in approval workflow, legal review triggers",
            "category": "OSS Approval",
        },
        {
            "id": "O-13",
            "requirement": "OSS alternatives evaluated before approval (build vs. buy vs. OSS decision)",
            "evidence": "Evaluation notes, decision records, alternative analysis documents",
            "category": "OSS Approval",
        },
        {
            "id": "O-14",
            "requirement": "OSS contribution guidelines exist if organisation contributes to OSS projects",
            "evidence": "Contribution policy, CLA (Contributor License Agreement) templates, approval process",
            "category": "OSS Contribution",
        },
        {
            "id": "O-15",
            "requirement": "OSS forks minimised and justified (upstream contributions preferred)",
            "evidence": "Fork justifications, upstream contribution plans, fork maintenance records",
            "category": "OSS Contribution",
        },
        {
            "id": "O-16",
            "requirement": "OSS components with known high-risk maintainers flagged for review",
            "evidence": "Maintainer reputation checks, risk assessments, red flag criteria",
            "category": "Supplier Risk",
        },
        {
            "id": "O-17",
            "requirement": "OSS inventory reviewed regularly (quarterly minimum) for accuracy",
            "evidence": "Inventory review reports, approval records, accuracy audit logs",
            "category": "OSS Metrics",
        },
        {
            "id": "O-18",
            "requirement": "OSS metrics tracked (% approved components, inventory coverage, SBOM completeness)",
            "evidence": "OSS compliance dashboard, metrics reports, KPI tracking",
            "category": "OSS Metrics",
        },
    ]

    return requirements

def get_dependency_vulnerability_requirements():
    """
    Domain 3: Dependency Vulnerability Management (18 requirements)

    Focus: SCA tools, vulnerability detection, remediation SLAs, lockfiles

    Returns:
        list: List of dicts with keys: id, requirement, evidence, category
    """
    requirements = [
        {
            "id": "D-01",
            "requirement": "SCA (Software Composition Analysis) tool deployed and scanning dependencies",
            "evidence": "SCA tool configurations (Snyk, Dependabot, WhiteSource, Sonatype), deployment records",
            "category": "SCA Tooling",
        },
        {
            "id": "D-02",
            "requirement": "SCA scans run automatically on every build/commit (not manual)",
            "evidence": "CI/CD pipeline configs, scan logs, automation evidence, build failure logs",
            "category": "SCA Tooling",
        },
        {
            "id": "D-03",
            "requirement": "SCA tool integrated with developer workflow (IDE plugins, PR checks)",
            "evidence": "IDE plugin configurations, PR check results, developer tool screenshots",
            "category": "SCA Tooling",
        },
        {
            "id": "D-04",
            "requirement": "Vulnerability alerts delivered to responsible teams (not ignored)",
            "evidence": "Alert routing configurations, notification logs, Slack/email integrations",
            "category": "Vulnerability Remediation",
        },
        {
            "id": "D-05",
            "requirement": "Vulnerability remediation SLAs defined by severity (Critical/High/Medium/Low)",
            "evidence": "SLA policy document, severity definitions, remediation timelines",
            "category": "Vulnerability Remediation",
        },
        {
            "id": "D-06",
            "requirement": "Critical vulnerabilities have <7 day remediation SLA (e.g., Log4Shell scenarios)",
            "evidence": "SLA policy with critical timelines, compliance tracking, escalation procedures",
            "category": "Vulnerability Remediation",
        },
        {
            "id": "D-07",
            "requirement": "High vulnerabilities have <30 day remediation SLA",
            "evidence": "SLA policy with high severity timelines, tracking metrics, compliance reports",
            "category": "Vulnerability Remediation",
        },
        {
            "id": "D-08",
            "requirement": "Vulnerability triage process documented (risk assessment, prioritization)",
            "evidence": "Triage workflow documentation, decision criteria, risk scoring methodology",
            "category": "Vulnerability Triage",
        },
        {
            "id": "D-09",
            "requirement": "False positives documented with suppression justification (not silently ignored)",
            "evidence": "Suppression records, justification notes, approval logs, suppression review process",
            "category": "Vulnerability Triage",
        },
        {
            "id": "D-10",
            "requirement": "Dependency update process documented (testing requirements, rollback plans)",
            "evidence": "Update procedures, testing requirements, rollback documentation, change management",
            "category": "Dependency Updates",
        },
        {
            "id": "D-11",
            "requirement": "Dependency updates tested before production deployment (not YOLO updates)",
            "evidence": "Test results, staging environment logs, regression test evidence, QA sign-offs",
            "category": "Dependency Updates",
        },
        {
            "id": "D-12",
            "requirement": "Transitive dependency vulnerabilities addressed (not just direct dependencies)",
            "evidence": "Transitive vulnerability remediation records, dependency tree analysis, patch logs",
            "category": "Dependency Management",
        },
        {
            "id": "D-13",
            "requirement": "Dependency pinning used (lockfiles required: package-lock.json, Pipfile.lock, etc.)",
            "evidence": "Lockfile examples, policy enforcement, CI/CD checks for lockfile presence",
            "category": "Dependency Management",
        },
        {
            "id": "D-14",
            "requirement": "Dependency version ranges minimised (avoid wildcards like '^1.0.0' or '~1.0.0')",
            "evidence": "Dependency manifests showing pinned versions, policy on version ranges",
            "category": "Dependency Management",
        },
        {
            "id": "D-15",
            "requirement": "Emergency patching process for critical vulnerabilities (Log4Shell-style response plan)",
            "evidence": "Emergency response plans, drill records, war room procedures, communication plans",
            "category": "Emergency Response",
        },
        {
            "id": "D-16",
            "requirement": "Vulnerability remediation tracked with metrics (MTTR, remediation rates)",
            "evidence": "Remediation dashboards, MTTR (Mean Time To Remediate) metrics, trend reports",
            "category": "Risk Governance",
        },
        {
            "id": "D-17",
            "requirement": "Unpatched vulnerabilities require formal risk acceptance (no silent ignoring)",
            "evidence": "Risk acceptance forms, approval signatures, risk register, acceptance expiry tracking",
            "category": "Risk Governance",
        },
        {
            "id": "D-18",
            "requirement": "SCA tool effectiveness reviewed regularly (false negative testing, coverage checks)",
            "evidence": "Tool effectiveness reports, benchmark tests, false negative rate analysis",
            "category": "Tool Effectiveness",
        },
    ]

    return requirements


def get_third_party_code_review_requirements():
    """
    Domain 4: Third-Party Code & Integration (18 requirements)

    Focus: Integration security, API assessment, package integrity, supply chain attacks

    Returns:
        list: List of dicts with keys: id, requirement, evidence, category
    """
    requirements = [
        {
            "id": "T-01",
            "requirement": "Third-party code integration policy exists (defines security review requirements)",
            "evidence": "Integration policy document, approval records, security review requirements",
            "category": "Integration Policy",
        },
        {
            "id": "T-02",
            "requirement": "Security review required for all third-party integrations before production",
            "evidence": "Review checklists, approval workflow, integration assessment records",
            "category": "Integration Policy",
        },
        {
            "id": "T-03",
            "requirement": "Third-party APIs assessed for security risks (authentication, authorisation, data exposure)",
            "evidence": "API security assessments, risk ratings, vulnerability scan results",
            "category": "API Security",
        },
        {
            "id": "T-04",
            "requirement": "API authentication mechanisms reviewed (OAuth, API keys, mutual TLS)",
            "evidence": "Authentication config reviews, security assessment notes, credential management",
            "category": "API Security",
        },
        {
            "id": "T-05",
            "requirement": "API authorisation controls verified (least privilege, scope limitations)",
            "evidence": "Authorisation testing results, access reviews, scope definitions, permission audits",
            "category": "API Security",
        },
        {
            "id": "T-06",
            "requirement": "Third-party API rate limiting and abuse prevention assessed",
            "evidence": "Rate limiting configurations, abuse detection logs, throttling evidence",
            "category": "API Security",
        },
        {
            "id": "T-07",
            "requirement": "SDK/library source code reviewed where feasible (especially for critical dependencies)",
            "evidence": "Code review reports, static analysis results, security findings",
            "category": "Code Review",
        },
        {
            "id": "T-08",
            "requirement": "Third-party container images scanned for vulnerabilities (Trivy, Clair, etc.)",
            "evidence": "Container scan reports, vulnerability findings, remediation evidence",
            "category": "Container Security",
        },
        {
            "id": "T-09",
            "requirement": "Container base images from trusted sources only (official registries, verified publishers)",
            "evidence": "Image provenance records, registry configurations, trusted source lists",
            "category": "Container Security",
        },
        {
            "id": "T-10",
            "requirement": "Package integrity verified before installation (checksums, GPG signatures)",
            "evidence": "Package verification configurations (GPG setup, SHA256 checks), verification logs",
            "category": "Package Integrity",
        },
        {
            "id": "T-11",
            "requirement": "Package manager security features enabled (npm audit, pip-audit, bundler-audit)",
            "evidence": "Package manager configurations, audit command integration, CI/CD security checks",
            "category": "Package Integrity",
        },
        {
            "id": "T-12",
            "requirement": "Typosquatting protection implemented (package name validation, allowlists)",
            "evidence": "Package allowlists, validation rules, namespace protection, monitoring alerts",
            "category": "Package Integrity",
        },
        {
            "id": "T-13",
            "requirement": "Third-party code isolated from critical systems (sandboxing, containerization)",
            "evidence": "Architecture diagrams, isolation configurations, namespace separation, security boundaries",
            "category": "Code Isolation",
        },
        {
            "id": "T-14",
            "requirement": "Vendor-provided code deployed with least privilege (minimal permissions)",
            "evidence": "Permission configurations, role definitions, privilege audits, access reviews",
            "category": "Code Isolation",
        },
        {
            "id": "T-15",
            "requirement": "Code escrow agreements in place for critical vendors (optional, for business continuity)",
            "evidence": "Escrow agreements, escrow deposit verification, code access procedures",
            "category": "Code Escrow",
        },
        {
            "id": "T-16",
            "requirement": "Third-party integration monitoring in place (API calls, error rates, anomalies)",
            "evidence": "Integration logs, monitoring dashboards, alert configurations, anomaly detection",
            "category": "Integration Monitoring",
        },
        {
            "id": "T-17",
            "requirement": "Supply chain attack mitigations documented (dependency confusion, typosquatting, compromised packages)",
            "evidence": "Mitigation strategies, security controls, attack surface analysis, prevention measures",
            "category": "Supply Chain Security",
        },
        {
            "id": "T-18",
            "requirement": "Third-party integration security reviewed regularly (annual reviews minimum)",
            "evidence": "Integration review reports, approval records, security reassessments, audit logs",
            "category": "Supply Chain Security",
        },
    ]

    return requirements


def get_license_compliance_requirements():
    """
    Domain 5: License Compliance & Legal Risk (18 requirements)

    Focus: License identification, GPL/copyleft management, legal compliance

    Returns:
        list: List of dicts with keys: id, requirement, evidence, category
    """
    requirements = [
        {
            "id": "L-01",
            "requirement": "License compliance policy exists and is approved by legal counsel",
            "evidence": "Policy document, legal approval signatures, policy distribution records",
            "category": "License Policy",
        },
        {
            "id": "L-02",
            "requirement": "License scanning tool deployed (FOSSA, Black Duck, WhiteSource, etc.)",
            "evidence": "License scan tool configurations, deployment records, integration evidence",
            "category": "License Scanning",
        },
        {
            "id": "L-03",
            "requirement": "License scans run automatically on every build (not manual/ad-hoc)",
            "evidence": "CI/CD configurations, scan logs, build pipeline evidence, automation scripts",
            "category": "License Scanning",
        },
        {
            "id": "L-04",
            "requirement": "All dependency licenses identified and tracked (no unknown licenses)",
            "evidence": "License scan reports, SBOM license data, license inventory, unknown license alerts",
            "category": "License Inventory",
        },
        {
            "id": "L-05",
            "requirement": "License allowlist and blocklist defined (approved vs. restricted licenses)",
            "evidence": "Allowlist/blocklist documents, approval criteria, legal guidance, license policies",
            "category": "License Inventory",
        },
        {
            "id": "L-06",
            "requirement": "Copyleft licenses flagged for legal review (GPL, AGPL, MPL, EPL)",
            "evidence": "Flagging rules in scan tools, legal review records, copyleft license alerts",
            "category": "Copyleft Management",
        },
        {
            "id": "L-07",
            "requirement": "License compatibility assessed (GPL + proprietary conflicts detected)",
            "evidence": "Compatibility matrix, conflict reports, license interaction analysis, legal opinions",
            "category": "Copyleft Management",
        },
        {
            "id": "L-08",
            "requirement": "Viral license contamination prevented (GPL/AGPL code isolated from proprietary code)",
            "evidence": "Isolation strategies, architecture reviews, code segregation, legal firewalls",
            "category": "Copyleft Management",
        },
        {
            "id": "L-09",
            "requirement": "Commercial licenses tracked with usage limits (seat counts, deployment restrictions)",
            "evidence": "License tracking system, usage reports, compliance monitoring, license renewals",
            "category": "License Enforcement",
        },
        {
            "id": "L-10",
            "requirement": "License violation detection automated (blocklist enforcement, conflict detection)",
            "evidence": "Violation detection rules, alert configurations, automated blocking, CI/CD gates",
            "category": "License Enforcement",
        },
        {
            "id": "L-11",
            "requirement": "License violations remediated with documented plan (not ignored)",
            "evidence": "Remediation plans, resolution records, alternative selection, legal consultation",
            "category": "License Enforcement",
        },
        {
            "id": "L-12",
            "requirement": "Legal review required for restrictive licenses (AGPL, SSPL, Business Source License)",
            "evidence": "Legal review requests, approval records, risk assessments, legal opinions",
            "category": "Legal Review",
        },
        {
            "id": "L-13",
            "requirement": "Open source attribution maintained (LICENSE, NOTICE, THIRD_PARTY_LICENSES files)",
            "evidence": "Attribution files in repositories, automated generation scripts, distribution packages",
            "category": "Attribution & Notices",
        },
        {
            "id": "L-14",
            "requirement": "Third-party license notices included in product distributions (legal requirement)",
            "evidence": "Distribution packages, license bundling, notice files, compliance documentation",
            "category": "Attribution & Notices",
        },
        {
            "id": "L-15",
            "requirement": "License compliance verified before production release (release gate)",
            "evidence": "Release checklists, compliance sign-offs, license approval records, gate evidence",
            "category": "Release Compliance",
        },
        {
            "id": "L-16",
            "requirement": "License audit trails maintained (license changes, approvals, violations)",
            "evidence": "Audit logs, compliance history, license change tracking, approval records",
            "category": "Release Compliance",
        },
        {
            "id": "L-17",
            "requirement": "License compliance training provided to developers (awareness of GPL, AGPL, etc.)",
            "evidence": "Training records, completion certificates, training materials, quiz results",
            "category": "License Metrics",
        },
        {
            "id": "L-18",
            "requirement": "License compliance metrics tracked (% compliant, violations, legal reviews)",
            "evidence": "Compliance dashboards, metrics reports, KPI tracking, trend analysis",
            "category": "License Metrics",
        },
    ]

    return requirements

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each assessment domain sheet sequentially.', '2. For each requirement, select Implementation Status from the dropdown menu.', '3. Provide specific evidence references (URLs, file paths, system names).', '4. Add Comments explaining implementation details or context.', '5. Focus on ACTUAL PRACTICES, not just documented policies.', '6. Review the Summary Dashboard to see overall compliance metrics.', '7. Document all gaps in the Gap Analysis sheet with remediation plans.', '8. Obtain required approvals on the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_domain_sheet(ws, domain_name, requirements, styles):
    """Create standardised assessment domain sheet matching gen 3 pattern."""
    thin = Side(style="thin")

    # Create a fresh DataValidation for implementation status on this sheet
    dv_status = DataValidation(
        type="list",
        formula1='"✅ Implemented,⚠️ Partially Implemented,❌ Not Implemented,N/A"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select a valid implementation status from the dropdown."
    )

    # Row 1: Title banner (merge A1:E1)
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Third-Party & OSS Assessment: {domain_name}".upper()
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO 27001:2022 | Control A.8.28 | {domain_name}"
    ws["A2"].font = Font(italic=True, size=10, color="003366", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Column headers (D9D9D9 fill, bold black font, thin border)
    headers = ["ID", "Requirement", "Implementation Status", "Evidence Reference", "Comments"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[3].height = 30
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40

    current_category = None
    row = 4

    for req in requirements:
        # Section banner when category changes
        if "category" in req and req["category"] != current_category:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"\u25b6 {req['category']}"
            ws[f"A{row}"].font = Font(bold=True, size=11, color="4472C4")
            ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            current_category = req["category"]

        # Column A: ID (bold, centred)
        ws[f"A{row}"] = req["id"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].font = Font(bold=True, size=9)

        # Column B: Requirement text (left, wrap, thin border)
        ws[f"B{row}"] = req["requirement"]
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Column C: Implementation Status (FFFFCC, dropdown, thin border)
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"C{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        dv_status.add(ws[f"C{row}"])

        # Column D: Evidence text (italic, grey, FFFFCC fill, thin border)
        ws[f"D{row}"] = req.get("evidence", "")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"D{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[f"D{row}"].font = Font(size=9, italic=True, color="666666")

        # Column E: Comments (FFFFCC fill, thin border)
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"E{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row += 1

    # Pad to 50 FFFFCC data rows for QA compliance (DS-005)
    min_data_rows = 50
    data_rows_written = len(requirements)
    if data_rows_written < min_data_rows:
        for pad_row in range(row, row + (min_data_rows - data_rows_written)):
            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}{pad_row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws[f"C{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"C{pad_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            dv_status.add(ws[f"C{pad_row}"])
            ws[f"D{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"D{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws[f"E{pad_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"E{pad_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Hide unused columns (F onwards)
    for col_letter in ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"
    if dv_status.sqref:
        ws.add_data_validation(dv_status)

    logger.info(f"\u2713 {domain_name} sheet created ({len(requirements)} requirements)")
    return ws

# ============================================================================
# SECTION 5: SUPPORTING SHEETS (SUMMARY, EVIDENCE, GAP, APPROVAL)
# ============================================================================

def create_summary_dashboard_sheet(wb, styles):
    """Create Summary Dashboard matching A.8.33 gold standard with TABLE 1, 2, 3."""
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False
    
    """Create Summary Dashboard matching A.8.33 gold standard with TABLE 1, 2, 3."""
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (italic, 003366 font, no fill)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 Annex A.8.28 — Secure Coding | Third-Party & OSS Assessment"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4: TABLE 1 banner (003366, merged A4:G4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=4, column=col).border = border

    # Row 5: Column headers (D9D9D9 grey, black text)
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows (4-8) - one per assessment area
    area_configs = [
        ("Vendor Security Assessment", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("OSS Management", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Dependency Vulnerability Mgmt", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("Third-Party Code & Integration", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
        ("License Compliance & Legal Risk", "C", ["✅ Implemented", "⚠️ Partially Implemented", "❌ Not Implemented"]),
    ]

    for i, (area_name, status_col, status_values) in enumerate(area_configs):
        row = 6 + i

        # Column A: Area name
        ws.cell(row=row, column=1, value=area_name).border = border

        # Column B: Total Items (COUNTA of status column)
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTA('{area_name}'!{status_col}4:{status_col}100)"
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal="center")

        # Column C: Compliant (✅ Implemented)
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[0]}")'
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal="center")

        # Column D: Partial (⚠️ Partially Implemented)
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[1]}")'
        cell_d.border = border
        cell_d.alignment = Alignment(horizontal="center")

        # Column E: Non-Compliant (❌ Not Implemented)
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"{status_values[2]}")'
        cell_e.border = border
        cell_e.alignment = Alignment(horizontal="center")

        # Column F: N/A count
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = f'=COUNTIF(\'{area_name}\'!{status_col}4:{status_col}100,"N/A")'
        cell_f.border = border
        cell_f.alignment = Alignment(horizontal="center")

        # Column G: Compliance % (Compliant / (Total - N/A))
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f'=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))'
        cell_g.number_format = "0.0%"
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center")

    # TOTAL row (after data rows + 2 buffer rows)
    total_row = 6 + len(area_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1).border = border
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    # Compliance % formula
    cell = ws.cell(row=total_row, column=7)
    cell.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    cell.number_format = "0.0%"
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

    # TABLE 2: Key Metrics
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{metrics_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=metrics_start, column=col).border = border

    # TABLE 2 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        hcell = ws.cell(row=metrics_start + 1, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center")

    # Metrics exploiting DVs from all 5 sheets
    metrics = [
        # Vendor Security
        ("Total Requirements Tracked", "=COUNTA('Vendor Security Assessment'!B5:B100)"),
        ("Security Requirements Documented", '=COUNTIF(\'Vendor Security Assessment\'!C5:C100,"✅ Implemented")'),
        ("Threat Modeling Coverage", '=COUNTIF(\'Vendor Security Assessment\'!C5:C100,"✅ Implemented")/COUNTA(\'Vendor Security Assessment\'!B5:B100)'),
        # OSS Management
        ("Total OSS Controls", "=COUNTA('OSS Management'!B5:B100)"),
        ("OSS Controls Implemented", '=COUNTIF(\'OSS Management\'!C5:C100,"✅ Implemented")'),
        ("OSS Repository Security Controls", '=COUNTIF(\'OSS Management\'!C5:C100,"✅ Implemented")'),
        # Dependency Vulnerability Management
        ("Total Pipeline Controls", "=COUNTA('Dependency Vulnerability Mgmt'!B5:B100)"),
        ("Security Scans Integrated", '=COUNTIF(\'Dependency Vulnerability Mgmt\'!C5:C100,"✅ Implemented")'),
        ("Deployment Gates Active", '=COUNTIF(\'Dependency Vulnerability Mgmt\'!C5:C100,"✅ Implemented")'),
        # Third-Party Code Review
        ("Total Testing Controls", "=COUNTA('Third-Party Code & Integration'!B5:B100)"),
        ("SAST/DAST Integration", '=COUNTIF(\'Third-Party Code & Integration\'!C5:C100,"✅ Implemented")'),
        ("Penetration Testing Coverage", '=COUNTIF(\'Third-Party Code & Integration\'!C5:C100,"✅ Implemented")'),
        # License Compliance
        ("Total License Controls", "=COUNTA('License Compliance & Legal Risk'!B5:B100)"),
        ("Approval Workflows Active", '=COUNTIF(\'License Compliance & Legal Risk\'!C5:C100,"✅ Implemented")'),
        ("License Management Compliance", '=COUNTIF(\'License Compliance & Legal Risk\'!C5:C100,"✅ Implemented")/COUNTA(\'License Compliance & Legal Risk\'!B5:B100)'),
    ]

    r = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=r, column=1, value=metric).border = border
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10)
        cell_val = ws.cell(row=r, column=2, value=formula)
        cell_val.border = border
        cell_val.font = Font(name="Calibri", size=10, color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # 2 buffer rows after TABLE 2 metrics (white, bordered)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # TABLE 3: Critical Findings (red banner)
    crit_start = r + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    # TABLE 3 column headers (D9D9D9, 7 cols)
    for col_idx, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Critical findings based on A.8.28 secure coding requirements
    findings = [
        ("Vendor Security Assessment", "Requirements without security acceptance criteria", '=COUNTIF(\'Vendor Security Assessment\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("OSS Management", "OSS controls not implemented", '=COUNTIF(\'OSS Management\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Dependency Vulnerability Mgmt", "Pipelines without SAST/SCA scanning", '=COUNTIF(\'Dependency Vulnerability Mgmt\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Third-Party Code & Integration", "Production code without security testing", '=COUNTIF(\'Third-Party Code & Integration\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("License Compliance & Legal Risk", "Releases without license approval", '=COUNTIF(\'License Compliance & Legal Risk\'!C5:C100,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Vendor Security Assessment", "Missing threat modeling for high-risk apps", '=COUNTIF(\'Vendor Security Assessment\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("OSS Management", "OSS controls partially implemented", '=COUNTIF(\'OSS Management\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Dependency Vulnerability Mgmt", "Security gates not enforced", '=COUNTIF(\'Dependency Vulnerability Mgmt\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("Third-Party Code & Integration", "Incomplete third-party code review coverage", '=COUNTIF(\'Third-Party Code & Integration\'!C5:C100,"⚠️ Partially Implemented")', "High", "Urgent"),
        ("License Compliance & Legal Risk", "License compliance gaps identified", '=COUNTIF(\'License Compliance & Legal Risk\'!C5:C100,"⚠️ Partially Implemented")', "Medium", "Plan"),
    ]

    r3 = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        ws.cell(row=r3, column=1, value=cat)
        ws.cell(row=r3, column=2, value=finding)
        count_cell = ws.cell(row=r3, column=3, value=formula)
        count_cell.alignment = Alignment(horizontal="center")
        count_cell.font = Font(name="Calibri", size=10, color="000000")
        ws.cell(row=r3, column=4, value=severity)
        ws.cell(row=r3, column=5, value=action)
        r3 += 1

    # 2 buffer rows after TABLE 3 (FFFFCC)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=r3, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=r3, column=col).border = border
        r3 += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"  # Freeze below subtitle + banner + headers

    return ws


# ============================================================================
# SECTION 7: EVIDENCE REGISTER CREATION
# ============================================================================

def create_evidence_register(wb, styles, validations):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border_thin
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Evidence Type dropdown
    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)

    # Verification Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)

    # MAX-001/MAX-002 fix: 1 grey sample + 100 FFFFCC empty rows
    # Sample row (row 5) with F2F2F2 grey background (like A.8.33 reference)
    sample_data = [
        "EV-001",
        "Test Data Inventory",
        "Configuration file",
        "Data masking configuration for UAT environment",
        "\\fileserver\\isms\\evidence\\test-data\\uat-masking-config.json",
        "2024-03-01",
        "Sarah Mitchell",
        "Verified"
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=5, column=col_num, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = border_thin

    # 100 empty rows (rows 6-105) - FFFFCC yellow
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
        dv_type.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"
    logger.info("Evidence Register sheet created")


def create_gap_analysis(wb, styles, validations):
    """Create Gap Analysis sheet for tracking remediation."""
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False

    # Column widths (10 columns to match WKBK 1-3)
    widths = {'A': 10, 'B': 22, 'C': 12, 'D': 35, 'E': 25, 'F': 25, 'G': 12, 'H': 20, 'I': 12, 'J': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Title (10 columns A-J)
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS & REMEDIATION PLAN"
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = styles['align_center']
    cell.border = styles['thick_border']
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Document all 'Not Implemented' and 'Partially Implemented' requirements. Assign owners, set target dates, and track remediation progress."
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = 3

    # Headers (standard blue 003366)
    headers = ["Gap ID", "Domain", "Requirement ID", "Requirement Description", "Current State",
               "Target State", "Priority", "Owner", "Target Date", "Status"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles['thin_border']

    # Hide unused columns (K onwards) to prevent empty column display
    for col_letter in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"
    
    # Add data validations
    ws.add_data_validation(validations['domain'])
    ws.add_data_validation(validations['priority'])
    ws.add_data_validation(validations['gap_status'])

    # MAX-003 fix: 1 sample row (F2F2F2 grey) + 50 empty rows (FFFFCC yellow) per Option B standard
    # Data starts at row 4 (row 1 = title, row 2 = subtitle, row 3 = headers)
    sample_row = 4
    sample_data = {
        'A': "GAP-001",
        'B': "Vendor Security Assessment",
        'C': "1.2",
        'D': "Third-party code review before integration",
        'E': "No security review of vendor code",
        'F': "Mandatory security review with SAST/dependency scan",
        'G': "High",
        'H': "Security Team Lead",
        'I': "28.02.2025",
        'J': "In Progress"
    }

    # Sample row with F2F2F2 grey fill
    for col, value in sample_data.items():
        ws[f'{col}{sample_row}'] = value
        ws[f'{col}{sample_row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws[f'{col}{sample_row}'].font = Font(name="Calibri", size=10, color="808080")
        ws[f'{col}{sample_row}'].border = styles['thin_border']

    # Apply validations to sample row
    validations['domain'].add(f'B{sample_row}')
    validations['priority'].add(f'G{sample_row}')
    validations['gap_status'].add(f'J{sample_row}')

    # 50 empty rows (FFFFCC yellow) - MAX-001/MAX-004 fix: no pre-filled IDs, 51 total rows
    for row in range(sample_row + 1, sample_row + 51):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{row}'].fill = styles['input_fill']
            ws[f'{col}{row}'].border = styles['thin_border']

        # Apply validations to empty rows
        validations['domain'].add(f'B{row}')
        validations['priority'].add(f'G{row}')
        validations['gap_status'].add(f'J{row}')

    logger.info("✓ Gap Analysis sheet created")


def create_approval_sheet(wb, styles, validations):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=1, column=col).border = border_thin
    ws.row_dimensions[1].height = 35

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=2, column=col).border = border_thin

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to all cells in merged range
    for col in range(1, 6):  # A:E
        ws.cell(row=row, column=col).border = border_thin

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.28.4 - Third-Party & OSS Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        # Apply borders to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = border_thin
            if value == "":
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Status dropdown on Assessment Status
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Apply borders to banner (A:E)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_thin
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            # Apply borders and fill to all cells in merged range (B:E)
            for col in range(2, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=row, column=col).border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all cells in merged range (B:E)
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border_thin

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to banner (A:E)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border_thin

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all cells in merged range (B:E)
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

    logger.info("\u2713 Approval Sign-Off sheet created")



def main():
    """Main execution function to generate the assessment workbook."""
    logger.info("=" * 70)
    logger.info("ISMS Control A.8.28.4 - Third-Party & OSS Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.28 (Secure Coding)")
    logger.info("=" * 70)
    logger.info("")
    
    # Initialize workbook
    logger.info("Initializing workbook...")
    wb = create_workbook()
    styles = get_style_definitions()
    validations = create_data_validations()
    column_widths = get_column_widths()
    logger.info("✓ Workbook initialized")
    logger.info("")
    
    # Create Instructions sheet
    logger.info("Creating Instructions sheet...")
    create_instructions_sheet(wb.create_sheet())
    logger.info("")
    
    # Create domain assessment sheets
    logger.info("Creating domain assessment sheets...")

    create_domain_sheet(
        wb.create_sheet("Vendor Security Assessment"),
        "Vendor Security Assessment",
        get_vendor_security_requirements(),
        styles
    )

    create_domain_sheet(
        wb.create_sheet("OSS Management"),
        "OSS Management",
        get_oss_management_requirements(),
        styles
    )

    create_domain_sheet(
        wb.create_sheet("Dependency Vulnerability Mgmt"),
        "Dependency Vulnerability Mgmt",
        get_dependency_vulnerability_requirements(),
        styles
    )

    create_domain_sheet(
        wb.create_sheet("Third-Party Code & Integration"),
        "Third-Party Code & Integration",
        get_third_party_code_review_requirements(),
        styles
    )

    create_domain_sheet(
        wb.create_sheet("License Compliance & Legal Risk"),
        "License Compliance & Legal Risk",
        get_license_compliance_requirements(),
        styles
    )
    
    logger.info("")
    
    # Create supporting sheets
    logger.info("Creating supporting sheets...")
    create_evidence_register(wb, styles, validations)
    create_gap_analysis(wb, styles, validations)
    create_summary_dashboard_sheet(wb, styles)
    create_approval_sheet(wb, styles, validations)
    logger.info("")
    
    # Save workbook
    logger.info("")
    logger.info("[S] Saving workbook...")
    filename = f"ISMS-IMP-A.8.28.4_Third_Party_OSS_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    # Summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("\u2705 SUCCESS: Third-Party & OSS Assessment workbook generated!")
    logger.info("=" * 70)
    logger.info("")
    logger.info(f"[D] File: {filename}")
    logger.info(f"[CHART] Total Requirements: 90 (18 per domain × 5 domains)")
    logger.info("")
    logger.info("Workbook Contents:")
    logger.info("  1. Instructions - Assessment guidance and methodology")
    logger.info("  2. Vendor Security Assessment - 18 requirements")
    logger.info("  3. OSS Management - 18 requirements")
    logger.info("  4. Dependency Vulnerability Mgmt - 18 requirements")
    logger.info("  5. Third-Party Code & Integration - 18 requirements")
    logger.info("  6. License Compliance & Legal Risk - 18 requirements")
    logger.info("  7. Summary Dashboard - Compliance metrics")
    logger.info("  8. Evidence Register - Evidence tracking")
    logger.info("  9. Gap Analysis - Remediation tracking")
    logger.info(" 10. Approval Sign-Off - Formal approval")
    logger.info("")
    logger.info("Total Requirements: 90 (18 per domain)")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Distribute workbook to assessment team")
    logger.info("  2. Complete domain assessments")
    logger.info("  3. Collect and register evidence")
    logger.info("  4. Track gaps and remediation")
    logger.info("  5. Obtain formal approvals")
    logger.info("")
    logger.info("Remember: Focus on what WORKS, not what's DOCUMENTED!")
    logger.info("=" * 70)
    
    return 0


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        logger.error(f"\n\u274C ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
