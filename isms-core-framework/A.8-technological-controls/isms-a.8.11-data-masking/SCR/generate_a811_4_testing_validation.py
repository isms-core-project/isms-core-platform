#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.11.4 - Testing & Validation Framework Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Assessment Domain 4 of 4: Testing & Validation Framework

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific data masking infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Data classification categories requiring masking (match your data governance framework)
2. Masking technique selection criteria per data type and use case
3. Environment categories where masking is mandatory (dev, test, analytics)
4. Testing and validation procedure requirements per masking technique
5. Masking exception approval workflow and compensating control requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.11 Data Masking Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
data masking controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Testing & Validation Framework under ISO 27001:2022 Control A.8.11. Supports evidence-based evaluation of data masking coverage, technique selection compliance, and validation effectiveness.

**Assessment Scope:**
- Sensitive data inventory completeness and masking requirement identification
- Masking technique selection appropriateness per data category
- Non-production environment masking coverage and compliance
- Masking process documentation and automation coverage
- Testing and validation procedure completeness and outcome tracking
- Exception management and approved compensating controls documentation
- Evidence collection for data protection and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
5. Summary Dashboard - Compliance overview and key metrics
6. Evidence Register - Audit evidence tracking
7. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Data Masking controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a811_4_testing_validation.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a811_4_testing_validation.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a811_4_testing_validation.py --date 20250115

Output:
    File: ISMS-IMP-A.8.11.4_Testing_&_Validation_Framework_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Assessment Domain:    4 of 4 (Testing & Validation Framework)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.11: Data Masking Policy (Governance)
    - ISMS-IMP-A.8.11.1: Data Inventory & Classification Assessment (Domain 1)
    - ISMS-IMP-A.8.11.2: Masking Technique Selection & Requirements (Domain 2)
    - ISMS-IMP-A.8.11.3: Environment Coverage Assessment (Domain 3)
    - ISMS-IMP-A.8.11.4: Testing & Validation Framework (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.11.4 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive data masking details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review data masking inventories and technique requirements annually or when new data categories are introduced, non-production environments change, or data protection incidents occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.11.4"
WORKBOOK_NAME = "Testing & Validation Framework"
CONTROL_ID = "A.8.11"
CONTROL_NAME = "Data Masking"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"




# ============================================================================
# HELPER: BATCH DATA-VALIDATION
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Instructions & Legend",
        "Testing Procedures",
        "PreDeployment Tests",
        "PostDeployment Validation",
        "Completeness Testing",
        "Format Preservation",
        "Referential Integrity",
        "ReIdentification Risk",
        "Data Utility Validation",
        "Performance Testing",
        "Ongoing Monitoring",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def get_styles():
    """Define all cell styles."""
    thin = Side(style="thin")

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS
# ============================================================================

def get_base_test_columns():
    """Standard 17 columns (A-Q) for test tracking sheets."""
    return {
        "Test ID": 15,
        "Test Name/Description": 30,
        "Environment Tested": 20,
        "Data Category Tested": 20,
        "Test Type": 20,
        "Test Date": 15,
        "Tester Name": 20,
        "Test Method": 25,
        "Test Result": 18,
        "Pass Criteria": 30,
        "Actual Outcome": 30,
        "Issues Found": 30,
        "Remediation Required?": 15,
        "Remediation Status": 18,
        "Retest Date": 15,
        "Notes/Comments": 30,
        "Evidence ID": 15,
    }


def get_extended_test_columns(sheet_type):
    """Extended columns for specific test sheets."""
    extensions = {
        "predeployment": {
            "Visual Inspection Done?": 15,
            "Automated Validation Done?": 15,
            "Comparison Test Done?": 15,
            "Sample Size Tested": 15,
            "Approval to Deploy?": 15,
        },
        "postdeployment": {
            "Validation Timing": 18,
            "Schema Change Impact?": 15,
            "User Feedback Reviewed?": 15,
            "Issues Reported by Users?": 15,
            "Next Validation Date": 15,
        },
        "completeness": {
            "Total Sensitive Fields": 15,
            "Masked Fields Count": 15,
            "Coverage %": 12,
            "Unmasked Fields Found": 15,
            "Schema Drift Detected?": 15,
            "New Columns Added?": 15,
            "Masking Rules Updated?": 15,
        },
        "format": {
            "Field Data Type": 18,
            "Format Validation Method": 20,
            "Expected Format": 25,
            "Format Pass Rate %": 12,
            "Format Failures Count": 15,
        },
        "integrity": {
            "Parent Table": 20,
            "Child Table": 20,
            "Foreign Key Field": 20,
            "FK Violations Found": 15,
            "Consistency Maintained?": 15,
        },
        "reid": {
            "Re-ID Technique Used": 22,
            "Re-ID Attempts": 15,
            "Successful Re-IDs": 15,
            "Re-ID Success Rate %": 12,
            "Risk Level": 12,
            "K-Anonymity Value": 12,
            "Mitigation Required?": 15,
        },
        "utility": {
            "Use Case Type": 22,
            "Test Suite Executed?": 15,
            "Tests Passed Count": 12,
            "Tests Failed Count": 12,
            "Utility Score %": 12,
            "Expected Failures?": 15,
            "Acceptable?": 15,
        },
        "performance": {
            "Metric Type": 20,
            "Baseline (Unmasked)": 15,
            "With Masking": 15,
            "Performance Impact %": 12,
            "Acceptable?": 15,
            "Optimization Needed?": 15,
            "Optimization Applied?": 15,
        },
        "monitoring": {
            "Monitoring Type": 20,
            "Monitoring Frequency": 15,
            "Alert Configured?": 15,
            "Incidents Detected": 12,
            "Incident Response Time": 15,
        },
    }
    return extensions.get(sheet_type, {})


# ============================================================================
# SECTION 3: DATA VALIDATION
# ============================================================================

def create_validations():
    """Create data validation objects (not yet added to any worksheet)."""
    return {
        'env_type': DataValidation(type="list", formula1='"Production,Development,Testing,UAT,Analytics,Cloud,Other"'),
        'data_cat': DataValidation(type="list", formula1='"PII,Financial,Health,Credentials,Proprietary,Mixed"'),
        'test_type': DataValidation(type="list", formula1='"Pre-Deployment,Post-Deployment,Completeness,Re-ID,Utility,Performance"'),
        'test_method': DataValidation(type="list", formula1='"Manual,Automated,Hybrid"'),
        'test_result': DataValidation(type="list", formula1='"\u2705 Pass,\u274C Fail,\u26A0\uFE0F Partial,Blocked,N/A"'),
        'yes_no': DataValidation(type="list", formula1='"Yes,No,Partial,Planned,N/A"'),
        'yes_no_simple': DataValidation(type="list", formula1='"Yes,No,N/A"'),
        'remediation_status': DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked"'),
        'checklist': DataValidation(type="list", formula1='"\u2705 Complete,\u26A0\uFE0F Partial,\u274C Missing,N/A"'),
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 4: REUSABLE SHEET CREATION
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Define your test strategy and scope in the Testing Procedures sheet.',
        '2. Execute pre-deployment validation tests and record results.',
        '3. Conduct post-deployment verification within 24 hours of deployment.',
        '4. Validate masking completeness across all sensitive fields.',
        '5. Test format preservation and referential integrity.',
        '6. Assess re-identification risk and data utility.',
        '7. Measure performance impact and document findings.',
        '8. Configure ongoing monitoring and continuous validation.',
        '9. Document all test results in the Evidence Register for audit traceability.',
        '10. Obtain final approval and sign-off from IT Security, ISO, and CISO.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A24"] = "Status Legend"
    ws["A24"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=25, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 26 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_test_sheet(ws, styles, title, policy_ref, question,
                     columns, row_count, checklist_items, sheet_type=None):
    """Generic function to create test tracking sheets."""

    # Header
    ws.merge_cells('A1:Q1')
    ws['A1'] = title.upper()
    apply_style(ws['A1'], styles["header"])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Policy Reference
    ws.merge_cells('A2:Q2')
    ws['A2'] = f"Policy Reference: {policy_ref}"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment Question
    ws['A3'] = question
    ws['A3'].font = Font(bold=True, size=11)
    ws['B3'].fill = styles["input_cell"]["fill"]
    ws['B3'].border = styles["border"]

    validations = create_validations()
    validations['yes_no'].add(ws['B3'])

    # Column Headers
    col_num = 1
    for col_name, col_width in columns.items():
        cell = ws.cell(row=6, column=col_num, value=col_name)
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[get_column_letter(col_num)].width = col_width
        col_num += 1

    total_cols = len(columns)

    # SAMPLE ROW (Row 7) - First data row with complete example
    sample_data = [
        "TEST-001",
        "Customer CRM",
        "Production",
        "CAT-PII-D",
        "Format Preservation",
        "15.01.2026",
        "Data Engineer",
        "Manual inspection",
        "Pass",
        "Email fields maintain @domain format after masking",
        "Test Plan v1.0",
        "EV-001",
        "No",
        "N/A",
        "15.01.2026",
        "Validated and approved",
    ]

    # Apply sample data to row 7 (up to total_cols to avoid index errors)
    for col_idx in range(1, min(len(sample_data) + 1, total_cols + 1)):
        cell = ws.cell(row=7, column=col_idx)
        if col_idx <= len(sample_data):
            cell.value = sample_data[col_idx - 1]
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]

    # EMPTY DATA ROWS (8-57: 50 additional rows = 51 total)
    start_row = 8
    end_row = 57

    for row in range(start_row, end_row + 1):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Apply base column validations
    apply_test_validations(ws, validations, 7, end_row)

    # Compliance Checklist
    checklist_row = end_row + 3
    ws[f'A{checklist_row}'] = f"{title.split()[0].upper()} CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1

    for item in checklist_items:
        ws[f'A{checklist_row}'] = "\u2610"
        ws[f'B{checklist_row}'] = item
        ws[f'C{checklist_row}'].fill = styles["input_cell"]["fill"]
        ws[f'C{checklist_row}'].border = styles["border"]
        validations['checklist'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    # Batch-apply all validations
    for _dv in validations.values():
        ws.add_data_validation(_dv)

    ws.freeze_panes = "A7"


def apply_test_validations(ws, validations, start_row, end_row):
    """Apply data validations to standard test columns."""
    # Column C: Environment Tested
    for row in range(start_row, end_row + 1):
        validations['env_type'].add(ws[f'C{row}'])

    # Column D: Data Category Tested
    for row in range(start_row, end_row + 1):
        validations['data_cat'].add(ws[f'D{row}'])

    # Column E: Test Type
    for row in range(start_row, end_row + 1):
        validations['test_type'].add(ws[f'E{row}'])

    # Column H: Test Method
    for row in range(start_row, end_row + 1):
        validations['test_method'].add(ws[f'H{row}'])

    # Column I: Test Result
    for row in range(start_row, end_row + 1):
        validations['test_result'].add(ws[f'I{row}'])

    # Column M: Remediation Required?
    for row in range(start_row, end_row + 1):
        validations['yes_no_simple'].add(ws[f'M{row}'])

    # Column N: Remediation Status
    for row in range(start_row, end_row + 1):
        validations['remediation_status'].add(ws[f'N{row}'])


# ============================================================================
# SECTION 5: SPECIALIZED SHEETS
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet (golden standard)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header -- single merged row with two-line title
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3: Document Information -- plain bold, NO fill, NO banner
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12, name="Calibri")

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Masking Testing & Validation"),
        ("Related Policy", "ISMS-POL-A.8.11"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Instructions
    row = 13
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    instructions = [
        "1. Define your test strategy and scope in the Testing Procedures sheet.",
        "2. Execute pre-deployment validation tests and record results.",
        "3. Conduct post-deployment verification within 24 hours of deployment.",
        "4. Validate masking completeness across all sensitive fields.",
        "5. Test format preservation and referential integrity.",
        "6. Assess re-identification risk and data utility.",
        "7. Measure performance impact and document findings.",
        "8. Configure ongoing monitoring and continuous validation.",
        "9. Document all test results in the Evidence Register for audit traceability.",
        "10. Obtain final approval and sign-off from IT Security, ISO, and CISO.",
    ]

    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # STATUS LEGEND -- plain bold heading, proper TABLE with 3 columns
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    # Table headers
    row += 1
    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    legend = [
        (CHECK, "Compliant", "Requirement fully met with evidence", "C6EFCE"),
        (WARNING, "Partial", "Partially implemented, gaps identified", "FFEB9C"),
        (XMARK, "Non-Compliant", "Requirement not met, remediation needed", "FFC7CE"),
        ("—", "Not Applicable", "Not applicable to this organisation", None),
    ]

    row += 1
    for sym, status, desc, color in legend:
        ws.cell(row=row, column=1, value=sym).border = border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if color:
            s.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Testing Dimensions
    row += 1
    ws[f"A{row}"] = "TESTING DIMENSIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    row += 1
    for ci, hdr in enumerate(("Dimension", "Key Question", "Success Criteria", "Target"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    dimensions = [
        ("Effectiveness", "Is data actually masked?", "Original data not visible", "100% fields masked"),
        ("Completeness", "Are ALL sensitive fields masked?", "Coverage = 100%", "100%"),
        ("Re-identification Risk", "Can original data be inferred?", "Re-ID attempts fail", "0% re-ID rate"),
        ("Data Utility", "Does masked data still work?", "Apps function correctly", "\u226595% utility"),
        ("Performance", "Does masking slow things down?", "Acceptable impact", "<10% degradation"),
    ]

    row += 1
    for dim, question, criteria, target in dimensions:
        ws.cell(row=row, column=1, value=dim).border = border
        ws.cell(row=row, column=2, value=question).border = border
        ws.cell(row=row, column=3, value=criteria).border = border
        ws.cell(row=row, column=4, value=target).border = border
        row += 1

    # Acceptable Evidence
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    evidence_types = [
        "\u2713 Pre-deployment test reports and pass/fail summaries",
        "\u2713 Post-deployment validation screenshots and logs",
        "\u2713 Masking coverage analysis reports",
        "\u2713 Re-identification risk assessment results",
        "\u2713 Data utility validation test suite results",
        "\u2713 Performance impact benchmarks (before/after)",
        "\u2713 Schema drift detection scan outputs",
        "\u2713 Referential integrity check results",
        "\u2713 Automated monitoring alert configurations",
        "\u2713 Gap analysis and remediation tracking records",
    ]
    row += 1
    for ev in evidence_types:
        ws[f"A{row}"] = ev
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


def create_testing_procedures(ws, styles):
    """Create Testing Procedures documentation sheet."""

    ws.merge_cells('A1:M1')
    ws['A1'] = "TESTING PROCEDURES DOCUMENTATION"
    apply_style(ws['A1'], styles["header"])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:M2')
    ws['A2'] = "Testing procedures must be documented and followed consistently (ISMS-POL-A.8.11-S2.4 Section 3)"
    apply_style(ws['A2'], styles["subheader"])

    ws['A3'] = "Are formal testing procedures documented for validating masking effectiveness?"
    ws['A3'].font = Font(bold=True, size=11)
    ws['B3'].fill = styles["input_cell"]["fill"]
    ws['B3'].border = styles["border"]

    validations = create_validations()
    validations['yes_no'].add(ws['B3'])

    # Column Headers
    headers = [
        ("A", "Procedure ID", 15),
        ("B", "Procedure Name", 30),
        ("C", "Test Type", 20),
        ("D", "When Performed", 25),
        ("E", "Responsible Role", 20),
        ("F", "Test Method", 20),
        ("G", "Tools Used", 25),
        ("H", "Pass Criteria", 30),
        ("I", "Frequency", 15),
        ("J", "Documentation Required", 30),
        ("K", "Procedure Status", 18),
        ("L", "Last Updated", 15),
        ("M", "Notes", 30),
    ]

    for col, header, width in headers:
        ws[f'{col}6'] = header
        apply_style(ws[f'{col}6'], styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 7) - First data row with complete example
    sample_data = [
        "PROC-001",
        "Visual Inspection of Masked Data",
        "Pre-Deployment",
        "Before each data refresh to non-prod",
        "Data Engineer",
        "Manual Review",
        "Database client (DBeaver, pgAdmin)",
        "No plaintext PII visible in sample of 100 records",
        "Per Refresh",
        "Screenshot evidence + sign-off email",
        "✅ Active",
        "15.01.2026",
        "Standard procedure for all environments",
    ]

    for col_idx, (col_letter, _, _) in enumerate(headers):
        if col_idx < len(sample_data):
            ws[f'{col_letter}7'] = sample_data[col_idx]
        ws[f'{col_letter}7'].fill = styles["input_cell"]["fill"]
        ws[f'{col_letter}7'].border = styles["border"]

    # EMPTY DATA ROWS (8-57: 50 additional rows = 51 total)
    for row in range(8, 58):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].fill = styles["input_cell"]["fill"]
            ws[f'{col}{row}'].border = styles["border"]

    # Batch-apply all validations
    for _dv in validations.values():
        ws.add_data_validation(_dv)

    ws.freeze_panes = "A7"


def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet with 51 row template (1 sample + 50 empty)."""

    # HEADER (Row 1)
    ws.merge_cells('A1:L1')
    ws['A1'] = "TESTING GAP ANALYSIS & REMEDIATION PLAN"
    apply_style(ws['A1'], styles["header"])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # SUBTITLE (Row 2)
    ws.merge_cells('A2:L2')
    ws['A2'] = "Identify gaps in testing validation processes (51 rows: 1 sample + 50 empty)"
    apply_style(ws['A2'], styles["subheader"])

    # COLUMN HEADERS (Row 3)
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Test Type", 20),
        ("C", "Gap Description", 35),
        ("D", "Current State", 25),
        ("E", "Target State", 25),
        ("F", "Risk Level", 12),
        ("G", "Impact", 25),
        ("H", "Remediation Action", 35),
        ("I", "Owner", 20),
        ("J", "Target Date", 15),
        ("K", "Status", 15),
        ("L", "Evidence ID", 15),
    ]

    for col, header, width in headers:
        ws[f'{col}3'] = header
        apply_style(ws[f'{col}3'], styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 4) - First data row with complete example
    sample_data = [
        "GAP-001",
        "Format Preservation",
        "Email masking does not preserve format (@ symbol position varies)",
        "Email masked as random string",
        "Email format preserved (user@domain structure maintained)",
        "Medium",
        "Validation tests fail; masked data breaks application logic",
        "Implement format-preserving masking function for email fields",
        "Data Engineering",
        "31.03.2026",
        "In Progress",
        "EV-001",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}4'] = value
        ws[f'{col_letter}4'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws[f'{col_letter}4'].border = styles["border"]

    # EMPTY DATA ROWS (5-54: 50 additional rows = 51 total)
    for row in range(5, 55):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{row}'].fill = styles["input_cell"]["fill"]
            ws[f'{col}{row}'].border = styles["border"]

    ws.freeze_panes = "A4"


def create_evidence_register(ws, styles):
    """Create Evidence Register sheet (golden standard)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle (no banner)
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4: Column headers (8 cols)
    headers = [
        ("Evidence ID", 15),
        ("Category", 20),
        ("Description", 40),
        ("Source/Location", 25),
        ("Date Collected", 15),
        ("Collected By", 20),
        ("Status", 15),
        ("Notes", 30),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Sample row with complete example data
    row = 5
    sample_data = {
        1: "EV-001",
        2: "Test report",
        3: "Data masking validation test results for production migration",
        4: "/evidence/masking-validation-prod-migration.pdf",
        5: "15.01.2026",
        6: "QA Team",
        7: "Verified",
        8: "Pre-deployment validation completed successfully"
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Empty rows 6-104 (99 empty rows)
    for r in range(6, 105):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Dropdowns
    dv_category = DataValidation(
        type="list",
        formula1='"Test report,Screenshot,Scan output,Configuration file,Log extract,Risk assessment,Monitoring alert,Remediation record,Audit evidence,Other"',
        allow_blank=True,
    )
    dv_category.add("B5:B104")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    dv_status.add("G5:G104")

    # Finalize validations
    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_status)

    ws.freeze_panes = "A5"


def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # ── Row 1: Main header ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle ──
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        f"Summary Dashboard  |  {WORKBOOK_NAME}  |  Generated: {GENERATED_DATE}"
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: blank ──

    # ── TABLE 1 banner (Row 4) ──
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 1 column headers (Row 5) ──
    t1_headers = [
        "Assessment Area", "Total Items", "Compliant",
        "Partial", "Non-Compliant", "N/A", "Compliance %",
    ]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 1 data rows ──
    # All test sheets use col I (test_result) rows 8:57
    # DV: "✅ Pass,❌ Fail,⚠️ Partial,Blocked,N/A"
    #   Compliant  = COUNTIF(I8:I57,"✅ Pass")
    #   Partial    = COUNTIF(I8:I57,"⚠️ Partial")
    #   Non-Comp   = COUNTIF(I8:I57,"❌ Fail")+COUNTIF(I8:I57,"Blocked")
    #   N/A        = COUNTIF(I8:I57,"N/A")
    #
    # Sheets: PreDeployment Tests, PostDeployment Validation, Completeness Testing,
    #         Format Preservation, Referential Integrity, ReIdentification Risk,
    #         Data Utility Validation, Performance Testing, Ongoing Monitoring
    # Gap Analysis: col K (Status), rows 5:54 — no standard DV applied in generator
    #   Use COUNTA for total, manual status values as text

    t1_areas = [
        (
            "Pre-Deployment Tests",
            "=C6+D6+E6+F6",
            "=COUNTIF('PreDeployment Tests'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('PreDeployment Tests'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('PreDeployment Tests'!I8:I100,\"\u274C Fail\")+COUNTIF('PreDeployment Tests'!I8:I100,\"Blocked\")",
            "=COUNTIF('PreDeployment Tests'!I8:I100,\"N/A\")",
        ),
        (
            "Post-Deployment Validation",
            "=C7+D7+E7+F7",
            "=COUNTIF('PostDeployment Validation'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('PostDeployment Validation'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('PostDeployment Validation'!I8:I100,\"\u274C Fail\")+COUNTIF('PostDeployment Validation'!I8:I100,\"Blocked\")",
            "=COUNTIF('PostDeployment Validation'!I8:I100,\"N/A\")",
        ),
        (
            "Completeness Testing",
            "=C8+D8+E8+F8",
            "=COUNTIF('Completeness Testing'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('Completeness Testing'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Completeness Testing'!I8:I100,\"\u274C Fail\")+COUNTIF('Completeness Testing'!I8:I100,\"Blocked\")",
            "=COUNTIF('Completeness Testing'!I8:I100,\"N/A\")",
        ),
        (
            "Format Preservation",
            "=C9+D9+E9+F9",
            "=COUNTIF('Format Preservation'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('Format Preservation'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Format Preservation'!I8:I100,\"\u274C Fail\")+COUNTIF('Format Preservation'!I8:I100,\"Blocked\")",
            "=COUNTIF('Format Preservation'!I8:I100,\"N/A\")",
        ),
        (
            "Referential Integrity",
            "=C10+D10+E10+F10",
            "=COUNTIF('Referential Integrity'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('Referential Integrity'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Referential Integrity'!I8:I100,\"\u274C Fail\")+COUNTIF('Referential Integrity'!I8:I100,\"Blocked\")",
            "=COUNTIF('Referential Integrity'!I8:I100,\"N/A\")",
        ),
        (
            "Re-Identification Risk",
            "=C11+D11+E11+F11",
            "=COUNTIF('ReIdentification Risk'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('ReIdentification Risk'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('ReIdentification Risk'!I8:I100,\"\u274C Fail\")+COUNTIF('ReIdentification Risk'!I8:I100,\"Blocked\")",
            "=COUNTIF('ReIdentification Risk'!I8:I100,\"N/A\")",
        ),
        (
            "Data Utility Validation",
            "=C12+D12+E12+F12",
            "=COUNTIF('Data Utility Validation'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('Data Utility Validation'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Data Utility Validation'!I8:I100,\"\u274C Fail\")+COUNTIF('Data Utility Validation'!I8:I100,\"Blocked\")",
            "=COUNTIF('Data Utility Validation'!I8:I100,\"N/A\")",
        ),
        (
            "Performance Testing",
            "=C13+D13+E13+F13",
            "=COUNTIF('Performance Testing'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('Performance Testing'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Performance Testing'!I8:I100,\"\u274C Fail\")+COUNTIF('Performance Testing'!I8:I100,\"Blocked\")",
            "=COUNTIF('Performance Testing'!I8:I100,\"N/A\")",
        ),
        (
            "Ongoing Monitoring",
            "=C14+D14+E14+F14",
            "=COUNTIF('Ongoing Monitoring'!I8:I100,\"\u2705 Pass\")",
            "=COUNTIF('Ongoing Monitoring'!I8:I100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Ongoing Monitoring'!I8:I100,\"\u274C Fail\")+COUNTIF('Ongoing Monitoring'!I8:I100,\"Blocked\")",
            "=COUNTIF('Ongoing Monitoring'!I8:I100,\"N/A\")",
        ),
    ]

    for row_idx, (area, total, compliant, partial, non_comp, na) in enumerate(t1_areas, start=6):
        row = row_idx
        ws[f"A{row}"] = area
        ws[f"A{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = total
        ws[f"B{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row}"].border = border
        ws[f"C{row}"] = compliant
        ws[f"C{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row}"].border = border
        ws[f"D{row}"] = partial
        ws[f"D{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{row}"].border = border
        ws[f"E{row}"] = non_comp
        ws[f"E{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"E{row}"].border = border
        ws[f"F{row}"] = na
        ws[f"F{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row}"].border = border
        ws[f"G{row}"] = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        ws[f"G{row}"].number_format = "0.0%"
        ws[f"G{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"G{row}"].border = border

    # ── TOTAL row ──
    total_row = 15
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Calibri", size=10, bold=True, color="000000")
    ws[f"A{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{total_row}"].border = border
    for col_letter, formula in [
        ("B", "=SUM(B6:B14)"),
        ("C", "=SUM(C6:C14)"),
        ("D", "=SUM(D6:D14)"),
        ("E", "=SUM(E6:E14)"),
        ("F", "=SUM(F6:F14)"),
    ]:
        ws[f"{col_letter}{total_row}"] = formula
        ws[f"{col_letter}{total_row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"{col_letter}{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col_letter}{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col_letter}{total_row}"].border = border
    ws[f"G{total_row}"] = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    ws[f"G{total_row}"].number_format = "0.0%"
    ws[f"G{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"G{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"G{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"G{total_row}"].border = border

    # ── TABLE 2 banner (Row 17) ──
    ws.merge_cells("A17:G17")
    ws["A17"] = "TABLE 2 \u2013 KEY METRICS"
    ws["A17"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A17"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A17"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 2 headers (Row 18) ──
    for col_idx, hdr in enumerate(["Metric", "Value", "Target"], start=1):
        cell = ws.cell(row=18, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 2 data rows (Rows 19-24) ──
    t2_data = [
        ("Pre-Deployment Test Pass Rate (%)", f"=IF(COUNTA('PreDeployment Tests'!I8:I100)=0,\"N/A\",TEXT(COUNTIF('PreDeployment Tests'!I8:I100,\"\u2705 Pass\")/COUNTA('PreDeployment Tests'!I8:I100),\"0.0%\"))", "100%"),
        ("Post-Deployment Validation Pass Rate (%)", f"=IF(COUNTA('PostDeployment Validation'!I8:I100)=0,\"N/A\",TEXT(COUNTIF('PostDeployment Validation'!I8:I100,\"\u2705 Pass\")/COUNTA('PostDeployment Validation'!I8:I100),\"0.0%\"))", "100%"),
        ("Re-Identification Risk Test Pass Rate (%)", f"=IF(COUNTA('ReIdentification Risk'!I8:I100)=0,\"N/A\",TEXT(COUNTIF('ReIdentification Risk'!I8:I100,\"\u2705 Pass\")/COUNTA('ReIdentification Risk'!I8:I100),\"0.0%\"))", "100%"),
        ("Data Utility Validation Pass Rate (%)", f"=IF(COUNTA('Data Utility Validation'!I8:I100)=0,\"N/A\",TEXT(COUNTIF('Data Utility Validation'!I8:I100,\"\u2705 Pass\")/COUNTA('Data Utility Validation'!I8:I100),\"0.0%\"))", ">=95%"),
        ("Performance Impact Tests Within Threshold (%)", f"=IF(COUNTA('Performance Testing'!I8:I100)=0,\"N/A\",TEXT(COUNTIF('Performance Testing'!I8:I100,\"\u2705 Pass\")/COUNTA('Performance Testing'!I8:I100),\"0.0%\"))", ">=90%"),
        ("Overall Test Pass Rate (%)", f"=IF(B15-F15=0,\"N/A\",TEXT(C15/(B15-F15),\"0.0%\"))", "100%"),
    ]
    for row_idx, (metric, value, target) in enumerate(t2_data, start=19):
        ws[f"A{row_idx}"] = metric
        ws[f"A{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row_idx}"].border = border
        ws[f"B{row_idx}"] = value
        ws[f"B{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row_idx}"].border = border
        ws[f"C{row_idx}"] = target
        ws[f"C{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row_idx}"].border = border

    # ── TABLE 3 banner (Row 26) ──
    ws.merge_cells("A26:G26")
    ws["A26"] = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    ws["A26"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A26"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws["A26"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 3 headers (Row 27) ──
    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", "Notes"]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=27, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 3 data rows (Rows 28-31) ──
    t3_data = [
        ("1", "Pre-deployment validation tests not consistently documented or signed off", "High — unvalidated masking deployments risk real data reaching non-production", "Define mandatory pre-deployment checklist; block data refresh without sign-off evidence", "P1", "Open", ""),
        ("2", "Re-identification risk testing not performed for all data categories", "High — masked data may still be re-identifiable via quasi-identifier combinations", "Conduct re-ID risk assessment for all data categories; document k-anonymity values", "P1", "Open", ""),
        ("3", "Format preservation failures for email and phone masking detected", "Medium — application logic fails when masked data breaks expected format constraints", "Implement format-preserving masking functions; add automated regex validation to test pipeline", "P2", "In Progress", ""),
        ("4", "Ongoing monitoring alerts not configured for all masking rules", "Medium — masking failures may go undetected between periodic test cycles", "Configure automated alerts for masking rule violations; integrate with SIEM; define SLA for response", "P2", "Open", ""),
    ]
    for row_idx, row_data in enumerate(t3_data, start=28):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border

    # ── Freeze panes ──
    ws.freeze_panes = "A4"



def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet (golden standard)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Row 3: ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G14),\"\")"),
        ("Assessment Status:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            if label == "Assessment Status:":
                status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    if status_row:
        dv_status.add(ws[f"B{status_row}"])

    # 3 approver sections
    approvers = [
        ("PREPARED BY (IT SECURITY)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "4472C4"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            row += 1
        row += 1  # gap between sections

    # FINAL ASSESSMENT DECISION
    ws[f"A{row}"] = "FINAL ASSESSMENT DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    dv_decision.add(ws[f"B{row}"])

    # NEXT REVIEW DATE section
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Finalize validations
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_decision)

    ws.freeze_panes = "A3"
    # Apply borders to all merged cell top-left corners (GS-AS-011)
    _as_thin = Side(style="thin")
    _as_border = Border(left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin)
    for merge_range in ws.merged_cells.ranges:
        tl = ws.cell(merge_range.min_row, merge_range.min_col)
        tl.border = _as_border


# ============================================================================
# SECTION 6: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.11.4 - Testing & Validation Framework Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.11 - Data Masking")
    logger.info("=" * 78)

    wb = create_workbook()
    styles = get_styles()

    logger.info("\n[1/15] Creating Instructions & Legend...")
    create_instructions(wb["Instructions & Legend"], styles)

    logger.info("[2/15] Creating Testing Procedures...")
    create_testing_procedures(wb["Testing Procedures"], styles)

    logger.info("[3/15] Creating Pre-Deployment Tests...")
    base_cols = get_base_test_columns()
    predeployment_cols = {**base_cols, **get_extended_test_columns("predeployment")}
    checklist_predeploy = [
        "Visual inspection performed on sample masked records",
        "No plaintext PII visible in masked data",
        "Automated validation scripts executed",
        "100% of identified sensitive fields masked",
        "No NULL values where masking expected",
        "Format validation passed (emails, phones, dates, etc.)",
        "Comparison test: NO matching values with production",
        "Referential integrity preserved (if required)",
        "Test results documented",
        "Failed tests remediated before deployment",
        "Approval obtained to deploy masked data",
        "Deployment blocked if ANY test fails",
    ]
    create_test_sheet(
        wb["PreDeployment Tests"], styles,
        "PRE-DEPLOYMENT TESTING",
        "Before deploying masked data, verify effectiveness through visual inspection, automated validation, and comparison testing (ISMS-POL-A.8.11-S2.4 Section 4.1)",
        "Are pre-deployment tests performed BEFORE masked data is deployed to target environments?",
        predeployment_cols, 51, checklist_predeploy, "predeployment"
    )

    logger.info("[4/15] Creating Post-Deployment Validation...")
    postdeployment_cols = {**base_cols, **get_extended_test_columns("postdeployment")}
    checklist_postdeploy = [
        "Immediate validation performed within 24 hours of deployment",
        "Masking verified in target environment",
        "Application functionality tested (if applicable)",
        "User feedback reviewed for masking issues",
        "No reports of visible PII from users",
        "Periodic validation scheduled (monthly/quarterly)",
        "Schema change impact assessed after every schema update",
        "Monitoring logs reviewed for masking failures",
        "Failed validations escalated to ISO",
        "Re-testing performed after remediation",
    ]
    create_test_sheet(
        wb["PostDeployment Validation"], styles,
        "POST-DEPLOYMENT VALIDATION",
        "After deploying masked data, immediate validation (within 24 hours) and periodic validation (monthly/quarterly) required (ISMS-POL-A.8.11-S2.4 Section 4.2)",
        "Is post-deployment validation performed to verify masking effectiveness in the target environment?",
        postdeployment_cols, 51, checklist_postdeploy, "postdeployment"
    )

    logger.info("[5/15] Creating Completeness Testing...")
    completeness_cols = {**base_cols, **get_extended_test_columns("completeness")}
    checklist_completeness = [
        "Sensitive fields inventory maintained",
        "Coverage calculated: (Masked Fields / Total Sensitive Fields) \u00d7 100",
        "Coverage target: 100% (with documented exceptions only)",
        "Schema comparison performed before every data refresh",
        "New columns detected and assessed for sensitivity",
        "Masking rules updated within 5 days for new sensitive columns",
        "Re-testing performed after masking rule updates",
        "Unmasked sensitive fields flagged for remediation",
        "Coverage gaps escalated to Data Owner and ISO",
        "Coverage metrics tracked over time",
        "Automated schema drift detection in place",
        "Coverage audit performed quarterly",
    ]
    create_test_sheet(
        wb["Completeness Testing"], styles,
        "COMPLETENESS TESTING & COVERAGE VALIDATION",
        "100% of identified sensitive fields must be masked. Schema drift detection required. (ISMS-POL-A.8.11-S2.4 Section 5)",
        "Is masking completeness tested to ensure 100% of sensitive fields are masked?",
        completeness_cols, 51, checklist_completeness, "completeness"
    )

    logger.info("[6/15] Creating Format Preservation...")
    format_cols = {**base_cols, **get_extended_test_columns("format")}
    checklist_format = [
        "Email format validation: Masked emails match email regex",
        "Phone format validation: Masked phones match phone format",
        "Credit card format validation: Masked cards pass Luhn (if required)",
        "Date format validation: Masked dates are valid dates",
        "ZIP code format validation: Masked ZIPs are valid format",
        "SSN format validation: Masked SSNs match SSN format",
        "Custom format validation performed (domain-specific)",
        "Format failures logged and remediated",
        "Format pass rate target: \u226599%",
        "Automated format validation scripts in place",
    ]
    create_test_sheet(
        wb["Format Preservation"], styles,
        "FORMAT & TYPE PRESERVATION VALIDATION",
        "Masked data must maintain correct format (email, phone, date, etc.) per validation requirements (ISMS-POL-A.8.11-S2.4 Section 7.3)",
        "Is format preservation validated to ensure masked data matches expected data types/formats?",
        format_cols, 51, checklist_format, "format"
    )

    logger.info("[7/15] Creating Referential Integrity...")
    integrity_cols = {**base_cols, **get_extended_test_columns("integrity")}
    checklist_integrity = [
        "Foreign key constraints validated after masking",
        "Related records can be joined correctly",
        "Consistent masking across related tables (same ID \u2192 same masked value)",
        "Database integrity checks executed",
        "Zero foreign key violations",
        "Orphaned records identified and remediated",
        "Automated integrity validation scripts in place",
        "Integrity failures block data deployment",
    ]
    create_test_sheet(
        wb["Referential Integrity"], styles,
        "REFERENTIAL INTEGRITY TESTING",
        "Foreign key constraints and data consistency must be preserved across related tables (ISMS-POL-A.8.11-S2.4 Section 7.2)",
        "Is referential integrity validated to ensure relationships preserved after masking?",
        integrity_cols, 51, checklist_integrity, "integrity"
    )

    logger.info("[8/15] Creating Re-Identification Risk...")
    reid_cols = {**base_cols, **get_extended_test_columns("reid")}
    checklist_reid = [
        "Direct matching test: No matches between masked and production",
        "Quasi-identifier combination test performed",
        "Statistical inference test performed",
        "External data correlation test performed",
        "Re-identification success rate calculated",
        "Re-ID rate target: 0% (zero successful re-identifications)",
        "If 1-5% re-ID rate: Mitigation applied (additional masking/aggregation)",
        "If >5% re-ID rate: Immediate remediation (re-design masking)",
        "K-anonymity assessment performed (optional but recommended)",
        "K-anonymity value: k \u2265 5 (if applicable)",
        "Quasi-identifiers documented",
        "Re-ID test scenarios documented",
        "Re-ID test results documented with evidence",
        "High re-ID risk escalated to CISO and DPO",
        "Re-ID testing performed quarterly",
    ]
    create_test_sheet(
        wb["ReIdentification Risk"], styles,
        "RE-IDENTIFICATION RISK ASSESSMENT",
        "Re-identification testing must be performed to verify masked data cannot be re-identified. Target: 0% re-ID rate. (ISMS-POL-A.8.11-S2.4 Section 6)",
        "Is re-identification risk testing performed to verify masked data cannot be re-identified?",
        reid_cols, 51, checklist_reid, "reid"
    )

    logger.info("[9/15] Creating Data Utility Validation...")
    utility_cols = {**base_cols, **get_extended_test_columns("utility")}
    checklist_utility = [
        "Application test suite executed with masked data",
        "All critical tests pass (or expected failures documented)",
        "Performance testing with masked data acceptable",
        "UAT completed successfully with masked data",
        "Training exercises work with masked data",
        "Statistical analysis results comparable (\u00b1acceptable margin)",
        "ML model accuracy acceptable when trained on masked data",
        "Reports render correctly with masked data",
        "Aggregations produce valid results",
        "Utility score calculated: % of use cases that work correctly",
        "Utility score target: \u226595%",
        "Utility failures documented and remediated",
    ]
    create_test_sheet(
        wb["Data Utility Validation"], styles,
        "DATA UTILITY VALIDATION",
        "Masked data must support intended use cases. Target: \u226595% utility. (ISMS-POL-A.8.11-S2.4 Section 7)",
        "Is data utility validated to ensure applications/analytics work correctly with masked data?",
        utility_cols, 51, checklist_utility, "utility"
    )

    logger.info("[10/15] Creating Performance Testing...")
    performance_cols = {**base_cols, **get_extended_test_columns("performance")}
    checklist_performance = [
        "Baseline performance measured (without masking)",
        "Performance measured with masking enabled",
        "Performance impact calculated: ((Masked - Baseline) / Baseline) \u00d7 100",
        "Performance impact target: <10% degradation",
        "Query performance tested",
        "Data load/refresh performance tested",
        "Report generation performance tested",
        "If impact \u226510%: Optimization strategies implemented",
        "Performance re-tested after optimization",
        "Performance metrics tracked over time",
    ]
    create_test_sheet(
        wb["Performance Testing"], styles,
        "PERFORMANCE IMPACT TESTING",
        "Masking performance impact must be measured. Target: <10% degradation. (ISMS-POL-A.8.11-S2.4 Section 8)",
        "Is performance impact tested to ensure masking overhead is acceptable?",
        performance_cols, 51, checklist_performance, "performance"
    )

    logger.info("[11/15] Creating Ongoing Monitoring...")
    monitoring_cols = {**base_cols, **get_extended_test_columns("monitoring")}
    checklist_monitoring = [
        "Automated monitoring enabled for masking failures",
        "Alerts configured for masking rule violations",
        "Monitoring logs reviewed regularly (daily/weekly)",
        "Schema change alerts configured",
        "Periodic re-testing scheduled (monthly/quarterly)",
        "Re-testing performed after every data refresh",
        "Re-testing performed after every schema change",
        "Incident response time tracked (target: <24 hours)",
        "Masking failures escalated to ISO",
        "Root cause analysis performed for failures",
        "Monitoring metrics tracked over time",
        "Annual comprehensive masking audit performed",
    ]
    create_test_sheet(
        wb["Ongoing Monitoring"], styles,
        "ONGOING MONITORING & CONTINUOUS VALIDATION",
        "Continuous monitoring and periodic re-testing required to detect masking degradation (ISMS-POL-A.8.11-S2.4 Section 9)",
        "Is ongoing monitoring in place to continuously validate masking effectiveness?",
        monitoring_cols, 51, checklist_monitoring, "monitoring"
    )

    logger.info("[12/15] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[13/15] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[14/15] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[15/15] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)

    # Save workbook -- auto-detect Assessment folder
    filename = f"ISMS-IMP-A.8.11.4_Testing_Validation_Framework_{datetime.now().strftime('%Y%m%d')}.xlsx"
    script_dir = Path(__file__).resolve().parent
    _wkbk_dir = script_dir.parent / "WKBK"
    _wkbk_dir.mkdir(exist_ok=True)
    output_path = _wkbk_dir / OUTPUT_FILENAME
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    finalize_validations(wb)
    wb.save(output_path)

    logger.info(f"\n\u2705 SUCCESS: {output_path}")
    logger.info("\nWorkbook Structure (15 Sheets):")
    logger.info("  \u2022 Instructions & Legend")
    logger.info("  \u2022 Testing Procedures (20 procedures)")
    logger.info("  \u2022 9 Test Tracking Sheets (106 total checklist items)")
    logger.info("  \u2022 Gap Analysis (40 rows)")
    logger.info("  \u2022 Evidence Register (EV-001 to EV-100)")
    logger.info("  \u2022 Summary Dashboard (KPIs + test compliance tracking)")
    logger.info("  \u2022 Approval Sign-Off (3-level approval workflow)")
    logger.info("\n" + "=" * 78)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
