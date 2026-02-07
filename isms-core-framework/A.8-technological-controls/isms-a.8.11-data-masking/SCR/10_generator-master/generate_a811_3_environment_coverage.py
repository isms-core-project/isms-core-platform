#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# Licensed under AGPL-3.0-or-later with commercial licensing option
#
# This file is part of the ISMS Compliance Framework
# See /LICENSE for full terms and /LICENSES/COMMERCIAL.md for commercial options
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.11.3 - Environment Coverage Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Assessment Domain 3 of 5: Environment Coverage & Implementation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific environment landscape and masking deployment
requirements.

Key customization areas:
1. Environment inventory (match your actual dev/test/prod environments)
2. Cloud provider options (AWS, Azure, GCP per your infrastructure)
3. Third-party sharing scenarios (specific to your data flows)
4. Backup and archive systems (aligned with your DR strategy)
5. Coverage thresholds (based on your risk appetite)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for data masking)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
masking implementation across production and non-production environments.

**Purpose:**
Enables systematic assessment of masking coverage across all environments
where sensitive data exists, ensuring comprehensive protection.

**Assessment Scope:**
- Production environment assessment
- Non-production environments (dev, test, QA, UAT, staging)
- Analytics and reporting environments
- Backup and archive environments
- External data sharing and third-party access
- Cloud and hybrid environment coverage
- Data flow mapping and coverage gaps
- Environment-specific masking requirements
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Environment assessment guidance
2. Environment Inventory - Complete environment catalog
3. Production Environment - Production masking assessment
4. Non-Production Environments - Dev/test/QA/UAT/staging coverage
5. Analytics & Reporting - BI and analytics environment coverage
6. Backup & Archive - Backup and archival data protection
7. External Sharing - Third-party and partner data sharing
8. Cloud Environments - Cloud and SaaS environment coverage
9. Data Flow Mapping - Cross-environment data flows
10. Gap Analysis - Coverage gaps and remediation
11. Evidence Register - Audit evidence tracking
12. Summary Dashboard - Environment coverage metrics

**Integration:**
This assessment feeds into ISMS-IMP-A.8.11.5 Compliance Dashboard, which
consolidates data from all five assessment domains for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a811_3_environment_coverage.py

Output:
    File: ISMS_IMP_A_8_11_3_Environment_Coverage_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Inventory all environments containing sensitive data
    2. Assess masking implementation in each environment
    3. Document environment-specific requirements
    4. Map data flows between environments
    5. Identify coverage gaps
    6. Define remediation actions
    7. Collect audit evidence
    8. Obtain stakeholder approvals
    9. Feed results into A.8.11.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Assessment Domain:    3 of 5 (Environment Coverage & Implementation)
Related Policy:       ISMS-POL-A.8.11 (Data Masking Policy)
Script Version:       1.0
Python Version:       3.8+

Related Documents:
    - ISMS-POL-A.8.11: Data Masking Policy
    - ISMS-IMP-A.8.11.1: Data Inventory & Classification (Domain 1)
    - ISMS-IMP-A.8.11.2: Masking Technique Selection (Domain 2)
    - ISMS-IMP-A.8.11.3: Environment Coverage Assessment Guide
    - ISMS-IMP-A.8.11.4: Testing & Validation (Domain 4)
    - ISMS-IMP-A.8.11.5: Compliance Dashboard (Consolidation)

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.11.3"
WORKBOOK_NAME = "Environment Coverage Assessment"
CONTROL_ID = "A.8.11"
CONTROL_NAME = "Data Masking"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Instructions_Legend",
        "Environment_Inventory",
        "Production_Environment",
        "NonProduction_Environments",
        "Analytics_Reporting",
        "Backup_Archive",
        "External_Sharing",
        "Cloud_Environments",
        "Data_Flow_Mapping",
        "Gap_Analysis",
        "Evidence_Register",
        "Summary_Dashboard",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def get_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS
# ============================================================================

def get_base_columns():
    """Standard 17 columns (A-Q) used across assessment sheets."""
    return {
        "Environment Name": 25,
        "Environment Type": 20,
        "Classification": 18,
        "Hosting Location": 18,
        "Data Sensitivity": 18,
        "Masking Required?": 18,
        "Masking Deployed?": 18,
        "Masking Technique": 20,
        "Masking Tool/Solution": 22,
        "Coverage %": 12,
        "Last Verified Date": 15,
        "Environment Owner": 20,
        "Data Owner": 20,
        "Exception Approved?": 15,
        "Compliance Status": 18,
        "Notes/Comments": 30,
        "Evidence ID": 15,
    }


def get_extended_columns(sheet_type):
    """Extended columns for specific sheets."""
    extensions = {
        "production": {
            "User Role/Group": 20,
            "Masked Fields": 25,
            "Unmasked Access Logged?": 15,
            "Access Control Method": 20,
            "Exception Justification": 30,
            "Risk Level": 12,
            "Remediation Target Date": 15,
        },
        "nonproduction": {
            "Data Refresh Frequency": 18,
            "Masking Applied During Refresh?": 18,
            "Direct Prod Clone Prevented?": 18,
            "Masking Validation Method": 22,
            "Last Data Refresh Date": 15,
            "Next Planned Refresh": 15,
            "Refresh Process Owner": 20,
        },
        "analytics": {
            "Analytics Platform Type": 22,
            "Aggregation Level": 18,
            "Re-ID Risk Assessed?": 15,
            "Re-ID Risk Level": 15,
            "Synthetic Data Used?": 15,
            "Data Export Controls": 22,
            "Self-Service BI Masking": 18,
        },
        "backup": {
            "Backup Type": 18,
            "Encryption Enabled?": 15,
            "Encryption Method": 20,
            "Access Control": 18,
            "Restoration Process": 22,
            "Masking on Restore?": 18,
            "Backup Retention Period": 15,
        },
        "external": {
            "Recipient Type": 20,
            "Data Sharing Purpose": 25,
            "DPA in Place?": 15,
            "DPA Specifies Masking?": 18,
            "Contractual Exception?": 18,
            "Recipient Security Audit Date": 15,
            "Data Minimization Applied?": 18,
        },
        "cloud": {
            "Cloud Provider": 18,
            "Cloud Service Type": 20,
            "Region/Geography": 18,
            "Client-Side Masking?": 18,
            "Cloud-Native Masking Tool": 22,
            "Multi-Tenancy Concerns?": 18,
            "Data Residency Compliance": 22,
        },
        "dataflow": {
            "Source Environment": 20,
            "Destination Environment": 20,
            "Data Type": 18,
            "Masking Checkpoint?": 18,
            "Masking Technique": 20,
            "Flow Frequency": 15,
            "Automated Masking?": 18,
            "Masking Tool/Script": 22,
            "Masking Validation": 20,
            "Last Flow Date": 15,
            "Flow Owner": 20,
            "Approval Required?": 15,
            "Approval Status": 18,
            "Compliance Status": 18,
            "Risk Level": 12,
            "Notes/Comments": 30,
            "Evidence ID": 15,
        },
    }
    return extensions.get(sheet_type, {})


# ============================================================================
# SECTION 3: DATA VALIDATION
# ============================================================================

def create_validations(ws):
    """Create and add data validation objects to worksheet."""
    validations = {
        'env_type': DataValidation(type="list", formula1='"Production,Development,Testing,UAT,Staging,Training,Sandbox,Analytics,Cloud,Backup,Archive,External"'),
        'classification': DataValidation(type="list", formula1='"Sensitive,Confidential,Internal,Public"'),
        'hosting': DataValidation(type="list", formula1='"On-Premises,AWS,Azure,GCP,Hybrid,Other Cloud"'),
        'data_sens': DataValidation(type="list", formula1='"PII,Financial,Health,Credentials,Proprietary,Mixed,None"'),
        'masking_req': DataValidation(type="list", formula1='"\u2705 Mandatory,\u26A0\uFE0F Conditional,\u274C Not Required,N/A"'),
        'yes_no': DataValidation(type="list", formula1='"Yes,No,Partial,Planned,N/A"'),
        'technique': DataValidation(type="list", formula1='"SDM,DDM,Tokenization,Encryption,Redaction,Substitution,Anonymization,None"'),
        'compliance': DataValidation(type="list", formula1='"\u2705 Compliant,\u26A0\uFE0F Partial,\u274C Non-Compliant,N/A"'),
        'yes_no_simple': DataValidation(type="list", formula1='"Yes,No,N/A"'),
        'checklist': DataValidation(type="list", formula1='"\u2705 Complete,\u26A0\uFE0F Partial,\u274C Missing,N/A"'),
    }
    
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    return validations


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 4: REUSABLE SHEET CREATION FUNCTIONS
# ============================================================================

def create_assessment_sheet(ws, styles, title, policy_ref, question, 
                           columns, row_count, checklist_items, sheet_type=None):
    """Generic function to create assessment sheets."""
    
    # Header
    ws.merge_cells('A1:Q1')
    ws['A1'] = title
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    # Policy Reference
    ws.merge_cells('A2:Q2')
    ws['A2'] = f"Policy Reference: {policy_ref}"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment Question
    ws['A3'] = question
    ws['A3'].font = Font(bold=True, size=11)
    ws['B3'].fill = styles["input_cell"]["fill"]
    
    validations = create_validations(ws)
    validations['yes_no'].add(ws['B3'])
    
    # Column Headers
    col_num = 1
    for col_name, col_width in columns.items():
        cell = ws.cell(row=6, column=col_num, value=col_name)
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[get_column_letter(col_num)].width = col_width
        col_num += 1
    
    total_cols = len(columns)
    
    # Data Entry Rows
    start_row = 8
    end_row = start_row + row_count - 1
    
    for row in range(start_row, end_row + 1):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
    
    # Apply specific validations based on sheet type
    apply_column_validations(ws, validations, start_row, end_row, sheet_type)
    
    # Compliance Checklist
    checklist_row = end_row + 3
    ws[f'A{checklist_row}'] = f"{title.split()[0]} CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    for item in checklist_items:
        ws[f'A{checklist_row}'] = "☐"
        ws[f'B{checklist_row}'] = item
        ws[f'C{checklist_row}'].fill = styles["input_cell"]["fill"]
        validations['checklist'].add(ws[f'C{checklist_row}'])
        checklist_row += 1
    
    ws.freeze_panes = "A7"


def apply_column_validations(ws, validations, start_row, end_row, sheet_type):
    """Apply data validations to specific columns."""
    
    # Column B: Environment Type
    for row in range(start_row, end_row + 1):
        validations['env_type'].add(ws[f'B{row}'])
    
    # Column C: Classification
    for row in range(start_row, end_row + 1):
        validations['classification'].add(ws[f'C{row}'])
    
    # Column D: Hosting Location
    for row in range(start_row, end_row + 1):
        validations['hosting'].add(ws[f'D{row}'])
    
    # Column E: Data Sensitivity
    for row in range(start_row, end_row + 1):
        validations['data_sens'].add(ws[f'E{row}'])
    
    # Column F: Masking Required?
    for row in range(start_row, end_row + 1):
        validations['masking_req'].add(ws[f'F{row}'])
    
    # Column G: Masking Deployed?
    for row in range(start_row, end_row + 1):
        validations['yes_no'].add(ws[f'G{row}'])
    
    # Column H: Masking Technique
    for row in range(start_row, end_row + 1):
        validations['technique'].add(ws[f'H{row}'])
    
    # Column N: Exception Approved?
    for row in range(start_row, end_row + 1):
        validations['yes_no_simple'].add(ws[f'N{row}'])
    
    # Column O: Compliance Status
    for row in range(start_row, end_row + 1):
        validations['compliance'].add(ws[f'O{row}'])


# ============================================================================
# SECTION 5: INSTRUCTIONS SHEET
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet."""
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "ISMS Control A.8.11.3 - Environment Coverage Assessment"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "ISO/IEC 27001:2022 - Data Masking Policy Compliance"
    apply_style(ws['A2'], styles["subheader"])
    
    # Document Info
    info = [
        ("Document ID:", "ISMS-IMP-A.8.11.3"),
        ("Assessment Area:", "Environment Coverage & Deployment"),
        ("Related Policy:", "ISMS-POL-A.8.11-S2.3"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row = 4
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "USER INPUT" in value:
            ws[f'B{row}'].fill = styles["input_cell"]["fill"]
        row += 1
    
    # How to Use
    row += 2
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    instructions = [
        "1. Complete each worksheet tab in sequence (Environment Inventory → Production → Non-Production → etc.)",
        "2. Fill ALL yellow-highlighted cells with your organization's specific information",
        "3. Use dropdown menus where provided (do not type free-form text in dropdown cells)",
        "4. Document ALL environments in your organization (include cloud, on-premises, hybrid)",
        "5. For each environment, specify masking requirement (Mandatory/Conditional/Not Required) per policy S2.3",
        "6. Verify masking deployment status (Yes/No/Partial/Planned)",
        "7. Calculate coverage percentage (% of sensitive fields masked)",
        "8. Link all assessments to Evidence Register with unique Evidence IDs",
        "9. Complete Gap Analysis sheet to identify remediation needs",
        "10. Review Summary Dashboard for executive-level compliance status",
    ]
    
    for instr in instructions:
        ws[f'A{row}'] = instr
        row += 1
    
    # Color Legend
    row += 2
    ws[f'A{row}'] = "COLOR LEGEND"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "🟨 Yellow cells = User input required"
    row += 1
    ws[f'A{row}'] = "🟩 Green status = Compliant (masking deployed as required)"
    row += 1
    ws[f'A{row}'] = "🟨 Yellow status = Partial compliance (some masking gaps)"
    row += 1
    ws[f'A{row}'] = "🟥 Red status = Non-compliant (masking required but not deployed)"
    row += 1
    ws[f'A{row}'] = "🟦 Blue status = Planned (remediation in progress)"
    
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 6: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet."""
    
    ws.merge_cells('A1:L1')
    ws['A1'] = "GAP ANALYSIS & REMEDIATION PLAN"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:L2')
    ws['A2'] = "All coverage gaps must be identified, risk-assessed, and remediated per organizational risk tolerance"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column Headers
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Environment/System", 25),
        ("C", "Gap Description", 35),
        ("D", "Current State", 25),
        ("E", "Target State", 25),
        ("F", "Risk Level", 12),
        ("G", "Impact", 25),
        ("H", "Remediation Action", 35),
        ("I", "Owner", 20),
        ("J", "Target Date", 15),
        ("K", "Status", 15),
        ("L", "Evidence ID", 15),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        apply_style(ws[f'{col}3'], styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data Rows
    validations = create_validations(ws)
    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked"')
    ws.add_data_validation(risk_dv)
    ws.add_data_validation(status_dv)
    
    for row in range(5, 45):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{row}'].fill = styles["input_cell"]["fill"]
            ws[f'{col}{row}'].border = styles["border"]
        
        risk_dv.add(ws[f'F{row}'])
        status_dv.add(ws[f'K{row}'])
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    
    ws.merge_cells('A1:J1')
    ws['A1'] = "EVIDENCE REGISTER"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:J2')
    ws['A2'] = "Supporting documentation for all environment coverage assessments"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column Headers
    headers = [
        ("A", "Evidence ID", 15),
        ("B", "Evidence Type", 20),
        ("C", "Description", 35),
        ("D", "Related Environment", 25),
        ("E", "Document Name/Link", 30),
        ("F", "Date Created", 15),
        ("G", "Owner", 20),
        ("H", "Retention Period", 15),
        ("I", "Location", 25),
        ("J", "Notes", 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        apply_style(ws[f'{col}3'], styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data Rows (100 rows)
    for row in range(5, 105):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{row}'].fill = styles["input_cell"]["fill"]
            ws[f'{col}{row}'].border = styles["border"]
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard sheet."""
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "ENVIRONMENT COVERAGE - EXECUTIVE SUMMARY DASHBOARD"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    # Compliance Summary Table
    ws['A3'] = "COMPLIANCE SUMMARY"
    ws['A3'].font = Font(bold=True, size=12)
    
    headers = ["Assessment Area", "Total Envs", "Compliant", "Partial", "Non-Compliant", "N/A", "Coverage %"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num, value=header)
        apply_style(cell, styles["column_header"])
    
    areas = [
        "Production Environments",
        "Non-Production Environments",
        "Analytics & Reporting",
        "Backup & Archive",
        "External Sharing",
        "Cloud Environments",
        "Data Flow Mapping",
        "OVERALL TOTAL",
    ]
    
    for row_num, area in enumerate(areas, start=5):
        ws[f'A{row_num}'] = area
        ws[f'A{row_num}'].font = Font(bold=(area == "OVERALL TOTAL"))
    
    # KPIs Section
    row = 16
    ws[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    kpi_headers = ["KPI", "Current Value", "Target", "Status"]
    for col_num, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_num, value=header)
        apply_style(cell, styles["column_header"])
    row += 1
    
    kpis = [
        ("Non-Production Masking Coverage", "100%"),
        ("Production DDM Coverage (role-based)", "≥90%"),
        ("Analytics Masking Coverage", "100%"),
        ("External Sharing Masking Coverage", "100%"),
        ("Cloud Environment Compliance", "100%"),
        ("Data Flow Mapping Completeness", "100%"),
        ("Exception Count (Total)", "≤5% of envs"),
        ("High-Risk Gaps", "0"),
        ("Average Remediation Time (days)", "≤90"),
        ("Evidence Documentation Rate", "100%"),
    ]
    
    for kpi, target in kpis:
        ws[f'A{row}'] = kpi
        ws[f'C{row}'] = target
        ws[f'B{row}'].fill = styles["input_cell"]["fill"]
        row += 1
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 9: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.11.3 - Environment Coverage Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.11 - Data Masking")
    logger.info("=" * 78)

    wb = create_workbook()
    styles = get_styles()
    
    logger.info("[1/12] Creating Instructions & Legend...")
    create_instructions(wb["Instructions_Legend"], styles)

    logger.info("[2/12] Creating Environment Inventory...")
    base_cols = get_base_columns()
    checklist_env = [
        "All production environments documented",
        "All non-production environments documented (Dev/Test/UAT/Sandbox)",
        "All analytics/reporting environments documented",
        "All cloud environments documented (AWS/Azure/GCP/Other)",
        "All backup/archive systems documented",
        "External data sharing destinations documented",
        "Each environment classified (Sensitive/Confidential/Internal/Public)",
        "Data sensitivity level assigned per environment",
        "Masking requirement determined (Mandatory/Conditional/Not Required)",
        "Environment owners assigned",
        "Data owners assigned",
        "Hosting location documented (On-Prem/Cloud)",
        "Environment inventory reviewed in last 6 months",
        "New environments added to inventory within 30 days of deployment",
        "Decommissioned environments removed from inventory",
    ]
    create_assessment_sheet(
    wb["Environment_Inventory"], styles,
        "ENVIRONMENT INVENTORY",
        "All information processing environments must be cataloged and classified for masking applicability (ISMS-POL-A.8.11-S2.3 Section 2)",
        "Does your organization maintain a complete inventory of ALL environments where data is processed, stored, or transmitted?",
    base_cols, 50, checklist_env
    )
    
    logger.info("[3/12] Creating Production Environment...")
    prod_cols = {**base_cols, **get_extended_columns("production")}
    checklist_prod = [
        "DDM implemented for role-based access in production",
        "Customer service representatives see masked customer data",
        "Production reports mask data for non-privileged users",
        "Audit logs containing sensitive data are masked/encrypted",
        "Application outputs (invoices, statements) use partial redaction",
        "Production data exports are masked before release",
        "All access to unmasked production data is logged",
        "Access logs reviewed monthly for anomalies",
        "Privileged user access to unmasked data requires justification",
        "Masking exceptions documented with business justification",
        "Exception approvals obtained from Data Owner and ISO",
        "Exceptions reviewed quarterly",
        "DDM performance impact assessed (<10% degradation acceptable)",
        "DDM rules tested before production deployment",
        "DDM bypass mechanisms disabled or strictly controlled",
        "Production masking coverage measured and tracked",
        "Coverage target: ≥90% of sensitive fields masked for non-privileged users",
        "Regulatory reporting requirements accommodate masking",
    ]
    create_assessment_sheet(
    wb["Production_Environment"], styles,
        "PRODUCTION ENVIRONMENT ASSESSMENT",
        "Production environments may use Dynamic Data Masking (DDM) for role-based access. All access to unmasked data must be logged. (ISMS-POL-A.8.11-S2.3 Section 3.1)",
        "Does your organization implement role-based masking (DDM) in production environments to restrict access to sensitive data?",
    prod_cols, 30, checklist_prod, "production"
    )
    
    logger.info("[4/12] Creating Non-Production Environments...")
    nonprod_cols = {**base_cols, **get_extended_columns("nonproduction")}
    checklist_nonprod = [
        "Development environments masked",
        "Testing/QA environments masked",
        "UAT environments masked",
        "Training environments masked",
        "Sandbox environments masked",
        "Staging environments masked (unless prod-identical required)",
        "Static Data Masking (SDM) applied during data refresh",
        "Masking applied BEFORE data deployment (not 'later')",
        "Direct production database cloning prevented",
        "Automated masking integrated into data refresh pipeline",
        "Manual data copies prohibited without masking",
        "Developer NDA reliance eliminated (NDAs not technical controls)",
        "Non-production data refresh process documented",
        "Masking validation performed after each refresh",
        "Non-production exception count: ___ (Target: 0)",
        "Exceptions approved by ISO and Data Owner",
        "Compensating controls implemented for exceptions",
        "Exception review conducted quarterly",
        "Coverage target: 100% of non-prod environments masked",
        "Non-compliance escalated to CISO",
    ]
    create_assessment_sheet(
    wb["NonProduction_Environments"], styles,
        "NON-PRODUCTION ENVIRONMENTS ASSESSMENT",
        "ALL sensitive data SHALL be masked before deployment to non-production. NO production data SHALL be copied without masking. (ISMS-POL-A.8.11-S2.3 Section 3.2)",
        "Are ALL non-production environments (Dev/Test/UAT/Training/Sandbox/Staging) masked before receiving production data?",
    nonprod_cols, 30, checklist_nonprod, "nonproduction"
    )
    
    logger.info("[5/12] Creating Analytics & Reporting...")
    analytics_cols = {**base_cols, **get_extended_columns("analytics")}
    checklist_analytics = [
        "BI tool exports contain masked data",
        "Data warehouse ETL includes masking steps",
        "Individual-level reports use masked data",
        "Aggregated reports assessed for re-identification risk",
        "ML/AI training datasets use synthetic or masked data",
        "PII removed or anonymized in training data",
        "Ad-hoc query exports masked",
        "Self-service BI tools enforce masking rules",
        "CSV/Excel exports contain masked data",
        "Analytics platform integrates masking at data ingestion",
        "Data minimization applied (only export necessary fields)",
        "Re-identification risk assessed before data release",
        "Historical analytics data masked or aggregated",
        "Analytics data retention period documented",
        "Analytics environment access logged",
    ]
    create_assessment_sheet(
    wb["Analytics_Reporting"], styles,
        "ANALYTICS & REPORTING ENVIRONMENTS ASSESSMENT",
        "Individual-level PII SHALL be masked or aggregated. Synthetic data SHALL be used for ML/AI training. Re-identification risk SHALL be assessed. (ISMS-POL-A.8.11-S2.3 Section 3.3)",
        "Are analytics, BI, data warehouse, and ML/AI environments masked to prevent individual-level data exposure?",
    analytics_cols, 30, checklist_analytics, "analytics"
    )
    
    logger.info("[6/12] Creating Backup & Archive...")
    backup_cols = {**base_cols, **get_extended_columns("backup")}
    checklist_backup = [
        "Production backups encrypted at rest",
        "Backup encryption uses strong algorithms (AES-256)",
        "Backup access restricted to authorized administrators only",
        "All backup access logged",
        "Non-production backups contain only masked data",
        "Backup restoration to non-prod triggers masking process",
        "Archive data assessed for masking feasibility",
        "If masking compromises compliance, encryption used instead",
        "Archive access strictly controlled",
        "Backup media stored securely (encrypted, access-controlled)",
        "Backup retention policy documented",
        "Backup testing includes masking validation",
    ]
    create_assessment_sheet(
    wb["Backup_Archive"], styles,
        "BACKUP & ARCHIVE ENVIRONMENTS ASSESSMENT",
        "Production backups may contain unmasked data if encrypted. Non-production backups SHALL contain only masked data. (ISMS-POL-A.8.11-S2.3 Section 3.4)",
        "Are backup and archive environments encrypted and access-controlled? Are non-production backups masked?",
    backup_cols, 30, checklist_backup, "backup"
    )
    
    logger.info("[7/12] Creating External Sharing...")
    external_cols = {**base_cols, **get_extended_columns("external")}
    checklist_external = [
        "Vendor data shares masked unless contractually required",
        "Data Processing Agreements (DPAs) in place",
        "DPAs specify masking requirements",
        "Vendor access to unmasked data logged and monitored",
        "Vendor environments audited for security controls",
        "Customer exports appropriately masked (not their own data)",
        "Auditor data samples masked where possible",
        "Auditors sign confidentiality agreements",
        "Partner data sharing follows minimum necessary principle",
        "Data minimization applied (only share necessary fields)",
        "External sharing inventory maintained",
        "Re-identification risk assessed before external release",
        "External sharing reviewed annually",
        "Data retention limits specified in agreements",
        "Data deletion verified after contract termination",
    ]
    create_assessment_sheet(
    wb["External_Sharing"], styles,
        "EXTERNAL DATA SHARING ASSESSMENT",
        "ALL data shared externally SHALL be masked unless contractually required. Data Processing Agreements SHALL specify masking. (ISMS-POL-A.8.11-S2.3 Section 3.5)",
        "Is data shared with vendors, partners, auditors, or customers masked unless contractually required to be unmasked?",
    external_cols, 30, checklist_external, "external"
    )
    
    logger.info("[8/12] Creating Cloud Environments...")
    cloud_cols = {**base_cols, **get_extended_columns("cloud")}
    checklist_cloud = [
        "Cloud environments classified correctly (Prod/Non-Prod/Analytics)",
        "Cloud-hosted production follows production masking rules",
        "Cloud-hosted non-prod follows non-prod masking rules (mandatory)",
        "Cloud analytics follows analytics masking rules",
        "Client-side masking applied before cloud upload (where applicable)",
        "Cloud provider security controls reviewed",
        "Cloud provider SLAs reviewed for data protection",
        "Multi-tenancy risks assessed and mitigated",
        "Data residency requirements met",
        "Cloud encryption enabled (at rest and in transit)",
        "Cloud access controls configured (IAM, RBAC)",
        "Cloud environment masking tested and validated",
    ]
    create_assessment_sheet(
    wb["Cloud_Environments"], styles,
        "CLOUD ENVIRONMENTS ASSESSMENT",
        "Cloud environments SHALL follow same masking requirements as on-premises. Cloud-hosted non-prod SHALL be masked. (ISMS-POL-A.8.11-S2.3 Section 2.6)",
        "Are cloud-hosted environments (AWS/Azure/GCP/Other) masked according to the same requirements as on-premises environments?",
    cloud_cols, 30, checklist_cloud, "cloud"
    )
    
    logger.info("[9/12] Creating Data Flow Mapping...")
    dataflow_cols = {
        "Data Flow Name": 25,
        "Source Environment": 20,
        "Destination Environment": 20,
        "Data Type": 18,
        "Masking Checkpoint?": 18,
        "Masking Technique": 20,
        "Flow Frequency": 15,
        "Automated Masking?": 18,
        "Masking Tool/Script": 22,
        "Masking Validation": 20,
        "Last Flow Date": 15,
        "Flow Owner": 20,
        "Approval Required?": 15,
        "Approval Status": 18,
        "Compliance Status": 18,
        "Risk Level": 12,
        "Notes/Comments": 30,
        "Evidence ID": 15,
    }
    checklist_dataflow = [
        "All production → non-production flows documented",
        "Masking checkpoints identified at environment boundaries",
        "Data refresh processes documented",
        "Automated masking integrated into data pipelines",
        "Manual data transfers prohibited or strictly controlled",
        "End-to-end masking verified for each flow",
        "Data flow inventory reviewed quarterly",
        "New data flows assessed for masking requirements",
        "Data flow exceptions approved by ISO",
        "High-risk flows (unmasked external sharing) escalated",
    ]
    create_assessment_sheet(
    wb["Data_Flow_Mapping"], styles,
        "DATA FLOW MAPPING & MASKING CHECKPOINTS",
        "Data flows SHALL be documented with masking checkpoints identified at each environment boundary. (ISMS-POL-A.8.11-S2.3 Section 3)",
        "Are data flows documented with masking checkpoints identified at each environment boundary?",
    dataflow_cols, 30, checklist_dataflow, "dataflow"
    )
    
    logger.info("[10/12] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap_Analysis"], styles)

    logger.info("[11/12] Creating Evidence Register...")
    create_evidence_register(wb["Evidence_Register"], styles)

    logger.info("[12/12] Creating Summary Dashboard...")
    create_summary_dashboard(wb["Summary_Dashboard"], styles)

    # Save workbook
    filename = f"ISMS-IMP-A.8.11.3_Environment_Coverage_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info("SUCCESS: %s", filename)
    logger.info("Workbook Structure: 12 sheets including 8 Assessment Sheets, Evidence Register")
    logger.info("=" * 78)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
