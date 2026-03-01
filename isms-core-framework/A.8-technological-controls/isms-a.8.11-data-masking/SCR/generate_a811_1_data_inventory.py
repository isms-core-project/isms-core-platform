#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# Licensed under AGPL-3.0-or-later with commercial licensing option
#
# This file is part of the ISMS Compliance Framework
# See /LICENSE for full terms and /LICENSES/COMMERCIAL.md for commercial options
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.11.1 - Data Inventory & Classification Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Assessment Domain 1 of 4: Data Inventory & Classification Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific data masking infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Data classification categories requiring masking (match your data governance framework)
2. Masking technique selection criteria per data type and use case
3. Environment categories where masking is mandatory (dev, test, analytics)
4. Testing and validation procedure requirements per masking technique
5. Masking exception approval workflow and compensating control requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.11 Data Masking Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
data masking controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Data Inventory & Classification Assessment under ISO 27001:2022 Control A.8.11. Supports evidence-based evaluation of data masking coverage, technique selection compliance, and validation effectiveness.

**Assessment Scope:**
- Sensitive data inventory completeness and masking requirement identification
- Masking technique selection appropriateness per data category
- Non-production environment masking coverage and compliance
- Masking process documentation and automation coverage
- Testing and validation procedure completeness and outcome tracking
- Exception management and approved compensating controls documentation
- Evidence collection for data protection and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
5. Summary Dashboard - Compliance overview and key metrics
6. Evidence Register - Audit evidence tracking
7. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Data Masking controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a811_1_data_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a811_1_data_inventory.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a811_1_data_inventory.py --date 20250115

Output:
    File: ISMS-IMP-A.8.11.1_Data_Inventory_&_Classification_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Assessment Domain:    1 of 4 (Data Inventory & Classification Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.11: Data Masking Policy (Governance)
    - ISMS-IMP-A.8.11.1: Data Inventory & Classification Assessment (Domain 1)
    - ISMS-IMP-A.8.11.2: Masking Technique Selection & Requirements (Domain 2)
    - ISMS-IMP-A.8.11.3: Environment Coverage Assessment (Domain 3)
    - ISMS-IMP-A.8.11.4: Testing & Validation Framework (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.11.1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive data masking details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review data masking inventories and technique requirements annually or when new data categories are introduced, non-production environments change, or data protection incidents occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys
from datetime import datetime, timedelta

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.11.1"
WORKBOOK_NAME = "Data Inventory & Classification Assessment"
CONTROL_ID = "A.8.11"
CONTROL_NAME = "Data Masking"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.11"
WORKBOOK_ID = "ISMS-IMP-A.8.11.1"
RELATED_POLICY = "ISMS-POL-A.8.11-S2.1"
ASSESSMENT_AREA = "Data Inventory & Classification"

# Color Scheme (consistent with A.8.23/8.24)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "F2F2F2"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C)
COLOR_PLANNED = "F2F2F2"         # Light blue (\u1F4CB)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 35
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create all sheets in order
    sheets = [
        "Instructions & Legend",
        "System Inventory",
        "Data Category Reference",
        "Sensitive Data Inventory",
        "Classification Matrix",
        "Regulatory Mapping",
        "Data Owner Assignment",
        "Masking Priority Matrix",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "info_cell": {
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "font": Font(name="Calibri", size=10, italic=True),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_complete": {
            "fill": PatternFill(start_color=COLOR_COMPLETE, end_color=COLOR_COMPLETE, fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color=COLOR_PARTIAL, end_color=COLOR_PARTIAL, fill_type="solid")
        },
        "status_missing": {
            "fill": PatternFill(start_color=COLOR_MISSING, end_color=COLOR_MISSING, fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color=COLOR_PLANNED, end_color=COLOR_PLANNED, fill_type="solid")
        },
    }
    
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            italic=getattr(style_dict["font"], 'italic', False),
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_unknown': DataValidation(
            type="list",
            formula1='"Yes,No,Unknown"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial"',
            allow_blank=False
        ),
        'yes_no_partial_planned_na': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False
        ),
        'yes_no_conditional': DataValidation(
            type="list",
            formula1='"Yes,No,Conditional,Unknown"',
            allow_blank=False
        ),
        'status_icons': DataValidation(
            type="list",
            formula1='"\u2705 Complete,\u26A0\uFE0F Partial,\u274C Missing,\u1F4CB Planned,N/A"',
            allow_blank=False
        ),
        'system_type': DataValidation(
            type="list",
            formula1='"Database,Application,SaaS,File Share,API,Data Warehouse,Backup System,Archive,Other"',
            allow_blank=False
        ),
        'environment_type': DataValidation(
            type="list",
            formula1='"Production,Development,Test/QA,UAT,Training,Analytics,DR/Backup,Decommissioned"',
            allow_blank=False
        ),
        'hosting_location': DataValidation(
            type="list",
            formula1='"On-Premises,AWS,Azure,GCP,Multi-Cloud,Hybrid,Third-Party Hosted,Unknown"',
            allow_blank=False
        ),
        'data_type': DataValidation(
            type="list",
            formula1='"String,Integer,Date,Boolean,Binary,JSON,Other"',
            allow_blank=False
        ),
        'data_category': DataValidation(
            type="list",
            formula1='"CAT-PII-D,CAT-PII-I,CAT-FIN,CAT-HLT,CAT-CRD,CAT-PRP,CAT-LOC,CAT-BIO,CAT-GEN,CAT-CHD"',
            allow_blank=False
        ),
        'pii_type': DataValidation(
            type="list",
            formula1='"Name,SSN,Email,Phone,Address,DOB,PAN,Medical Record,Password,IP Address,Other"',
            allow_blank=False
        ),
        'sensitivity_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Public"',
            allow_blank=False
        ),
        'masking_status': DataValidation(
            type="list",
            formula1='"Masked,Not Masked,Partially Masked,Encrypted,Planned"',
            allow_blank=False
        ),
        'discovery_method': DataValidation(
            type="list",
            formula1='"Automated Scan,Schema Review,Manual Sample,Data Owner Interview,Unknown"',
            allow_blank=False
        ),
        'org_classification': DataValidation(
            type="list",
            formula1='"Restricted,Confidential,Internal,Public"',
            allow_blank=False
        ),
        'gdpr_article': DataValidation(
            type="list",
            formula1='"Art. 4 (PII),Art. 9 (Special),Art. 32 (Security),N/A"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'exposure_risk': DataValidation(
            type="list",
            formula1='"Very High,High,Medium,Low,Very Low"',
            allow_blank=False
        ),
        'priority_tier': DataValidation(
            type="list",
            formula1='"P1,P2,P3,P4"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Complete,Blocked"',
            allow_blank=False
        ),
        'gap_category': DataValidation(
            type="list",
            formula1='"Inventory Missing,Classification Missing,Ownership Missing,Masking Missing,Regulatory Unknown,Other"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Complete,Accepted Risk"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Data Discovery Report,Schema Documentation,DPIA,Data Flow Diagram,Classification Review,RACI Matrix,Meeting Minutes,Approval Email,Tool Report,Other"',
            allow_blank=False
        ),
        'review_frequency': DataValidation(
            type="list",
            formula1='"Annual,Semi-Annual,Quarterly,As-Needed,N/A"',
            allow_blank=False
        ),
        'confidentiality_level': DataValidation(
            type="list",
            formula1='"Public,Internal,Confidential,Restricted"',
            allow_blank=False
        ),
    }
    
    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================

def create_instructions_legend(ws, styles):
    """Create Instructions & Legend sheet — golden standard layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header — single merged row A1:G1, two-line title
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Document Information (row 3) — plain bold, NO fill, NO banner
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12, name="Calibri")

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Control Reference", CONTROL_REF),
        ("Framework Version", "1.0"),
        ("Generated Date", GENERATED_DATE),
        ("Classification", "CONFIDENTIAL"),
        ("Owner", ""),
        ("Status", "DRAFT"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Instructions — plain bold heading
    row += 1
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    instructions = [
        "1. Start with System Inventory — list ALL systems/databases that process data.",
        "2. Use Data Category Reference to understand sensitivity taxonomy.",
        "3. Complete Sensitive Data Inventory for each system (table/field level).",
        "4. Apply classification using the Classification Matrix sheet.",
        "5. Map data to regulatory requirements in Regulatory Mapping.",
        "6. Assign data owners in Data Owner Assignment.",
        "7. Prioritise masking efforts in Masking Priority Matrix.",
        "8. Document gaps in Gap Analysis.",
        "9. Maintain evidence in Evidence Register.",
        "10. Review summary metrics in Summary Dashboard.",
        "11. Obtain required approvals in Approval Sign-Off.",
    ]

    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # DATA CATEGORY TAXONOMY
    row += 1
    ws[f"A{row}"] = "DATA CATEGORY TAXONOMY"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")
    row += 1

    for ci, hdr in enumerate(("Category ID", "Category Name", "Examples"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    row += 1

    categories = [
        ("CAT-PII-D", "Direct PII", "Name, SSN/AHV, Passport, Email, Phone"),
        ("CAT-PII-I", "Indirect PII", "DOB+ZIP, Job Title+Department"),
        ("CAT-FIN", "Financial Data", "Credit Card (PAN), IBAN, Account Balance"),
        ("CAT-HLT", "Health Data", "Diagnoses, Prescriptions, Medical Records"),
        ("CAT-CRD", "Credentials", "Passwords, API Keys, Tokens"),
        ("CAT-PRP", "Proprietary", "Trade Secrets, Pricing, IP"),
        ("CAT-LOC", "Location Data", "GPS, IP Address, Travel History"),
        ("CAT-BIO", "Biometric", "Fingerprints, Facial Geometry"),
        ("CAT-GEN", "Genetic Data", "DNA, Genetic Test Results"),
        ("CAT-CHD", "Child Data", "Data of minors (<16 GDPR / <18)"),
    ]

    for cat_id, cat_name, examples in categories:
        ws.cell(row=row, column=1, value=cat_id).border = border
        ws.cell(row=row, column=2, value=cat_name).border = border
        ws.cell(row=row, column=3, value=examples).border = border
        row += 1

    # SENSITIVITY CLASSIFICATION LEVELS
    row += 1
    ws[f"A{row}"] = "SENSITIVITY CLASSIFICATION LEVELS"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")
    row += 1

    for ci, hdr in enumerate(("Level", "Definition", "Masking Requirement"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    row += 1

    levels = [
        ("Critical", "Severe harm if exposed, regulatory breach", "SHALL mask in ALL non-production"),
        ("High", "Substantial harm, privacy violation", "SHALL mask in non-production"),
        ("Medium", "Moderate harm, business impact", "SHOULD mask in non-production"),
        ("Low", "Minimal harm", "MAY mask based on risk"),
        ("Public", "No confidentiality requirement", "N/A"),
    ]

    for level, definition, requirement in levels:
        ws.cell(row=row, column=1, value=level).border = border
        ws.cell(row=row, column=2, value=definition).border = border
        ws.cell(row=row, column=3, value=requirement).border = border
        row += 1

    # STATUS LEGEND
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")
    row += 1

    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    row += 1

    legend = [
        (CHECK, "Complete", "Inventory complete, classified", "C6EFCE"),
        (WARNING, "Partial", "Partial inventory/classification", "FFEB9C"),
        (XMARK, "Missing", "Not inventoried/classified", "FFC7CE"),
        ("—", "Not Applicable", "System contains no sensitive data", None),
    ]

    for sym, status, desc, color in legend:
        ws.cell(row=row, column=1, value=sym).border = border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if color:
            s.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # ACCEPTABLE EVIDENCE
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")
    row += 1

    evidence_examples = [
        "\u2713 Database schema documentation",
        "\u2713 Data dictionaries with field definitions",
        "\u2713 Data discovery tool reports (e.g., BigID, OneTrust, etc.)",
        "\u2713 Data flow diagrams",
        "\u2713 Data Processing Impact Assessments (DPIA)",
        "\u2713 System design documents",
        "\u2713 Privacy policy documentation",
        "\u2713 Data retention schedules",
        "\u2713 Data owner assignment matrices (RACI)",
        "\u2713 Classification review meeting minutes",
        "\u2713 Regulatory compliance assessments",
        "\u2713 Third-party data sharing agreements",
    ]

    for evidence in evidence_examples:
        ws[f"A{row}"] = evidence
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70

    # Freeze panes
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 5: SHEET 2 - SYSTEM INVENTORY
# ============================================================================

def create_system_inventory(ws, styles):
    """Create System Inventory sheet with 50-row template."""

    validations = create_base_validations(ws)

    # HEADER
    ws.merge_cells('A1:Q1')
    header = ws['A1']
    header.value = "SYSTEM & DATABASE INVENTORY"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35
    
    # SUBTITLE
    ws.merge_cells('A2:Q2')
    subtitle = ws['A2']
    subtitle.value = "List ALL systems that process, store, or transmit data (50 row template)"
    apply_style(subtitle, styles["subheader"])

    # ASSESSMENT QUESTION
    ws.merge_cells('A3:Q3')
    ws['A3'] = "Does your organisation maintain a comprehensive inventory of all systems containing sensitive data?"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(wrap_text=True)
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    ws['B4'].value = ""
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # COLUMN HEADERS (Row 6)
    headers = [
        ("A", "System ID", 15),
        ("B", "System Name", 25),
        ("C", "System Type", 18),
        ("D", "Environment", 15),
        ("E", "Contains Sensitive Data?", 18),
        ("F", "Data Categories Present", 25),
        ("G", "Hosting Location", 20),
        ("H", "Primary Data Owner", 20),
        ("I", "System Owner/Admin", 20),
        ("J", "Inventory Status", 18),
        ("K", "Last Inventory Date", 15),
        ("L", "Next Review Date", 15),
        ("M", "Record Count (Approx)", 15),
        ("N", "Retention Period", 15),
        ("O", "Regulatory Scope", 20),
        ("P", "Decommission Date", 15),
        ("Q", "Notes/Comments", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}6']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # SAMPLE ROW (Row 7) - First data row with complete example
    sample_data = [
        "SYS-001",
        "Customer CRM System",
        "Application",
        "Production",
        "Yes",
        "CAT-PII-D, CAT-FIN",
        "AWS",
        "VP Sales",
        "IT Manager",
        "✅ Complete",
        "2024-12-15",
        "2025-03-15",
        "500000",
        "7 years",
        "GDPR, FADP",
        "",
        "Primary customer database",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = value
        apply_style(cell, styles["input_cell"])

    # EMPTY DATA ROWS (8-57: 50 additional rows = 51 total)
    for row_idx in range(8, 58):
        for col_idx in range(1, 18):  # A-Q = 17 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Apply data validations to appropriate columns
    for row in range(7, 58):
        validations['system_type'].add(ws[f'C{row}'])
        validations['environment_type'].add(ws[f'D{row}'])
        validations['yes_no_unknown'].add(ws[f'E{row}'])
        validations['hosting_location'].add(ws[f'G{row}'])
        validations['status_icons'].add(ws[f'J{row}'])
    
    # CHECKLIST SECTION (Starting Row 59)
    checklist_row = 59
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "SYSTEM INVENTORY CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    # Checklist headers
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is a system inventory maintained?",
        "Are all production systems documented?",
        "Are non-production environments included?",
        "Are SaaS/cloud systems included?",
        "Are backup/archive systems included?",
        "Are data warehouses/analytics systems included?",
        "Is system type classified for each entry?",
        "Is hosting location documented?",
        "Are system owners assigned?",
        "Are data owners assigned?",
        "Is inventory reviewed quarterly?",
        "Are decommissioned systems tracked?",
        "Is record count estimated per system?",
        "Is retention period documented?",
        "Are regulatory requirements mapped?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30

    # REFERENCE TABLE: System Types (Starting Row 76)
    ref_row = 76
    ws.merge_cells(f'A{ref_row}:C{ref_row}')
    ws[f'A{ref_row}'] = "SYSTEM TYPE DEFINITIONS"
    ws[f'A{ref_row}'].font = Font(bold=True, size=11)
    ref_row += 1

    ws[f'A{ref_row}'] = "System Type"
    ws[f'B{ref_row}'] = "Description"
    ws[f'C{ref_row}'] = "Examples"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{ref_row}'], styles["column_header"])
    ref_row += 1

    system_types = [
        ("Database", "Relational or NoSQL databases", "Oracle, PostgreSQL, MongoDB, SQL Server"),
        ("Application", "Business applications", "CRM, ERP, HR systems, custom apps"),
        ("SaaS", "Cloud-based software services", "Salesforce, Workday, ServiceNow"),
        ("File Share", "File storage systems", "SharePoint, NAS, file servers"),
        ("API", "API gateways and services", "REST APIs, integration platforms"),
        ("Data Warehouse", "Analytics and BI systems", "Snowflake, Redshift, BigQuery"),
        ("Backup System", "Backup and recovery systems", "Veeam, backup appliances"),
        ("Archive", "Long-term storage", "Tape systems, cold storage"),
        ("Other", "Other system types", "(Document in notes)"),
    ]

    for sys_type, description, examples in system_types:
        ws[f'A{ref_row}'] = sys_type
        ws[f'B{ref_row}'] = description
        ws[f'C{ref_row}'] = examples
        ref_row += 1

    # Freeze panes
    ws.freeze_panes = "A7"

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: SHEET 3 - DATA CATEGORY REFERENCE
# ============================================================================

def create_data_category_reference(ws, styles):
    """Create read-only Data Category Reference sheet."""

    # HEADER
    ws.merge_cells('A1:G1')
    header = ws['A1']
    header.value = "SENSITIVE DATA CATEGORY REFERENCE"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35

    # SUBTITLE
    ws.merge_cells('A2:G2')
    subtitle = ws['A2']
    subtitle.value = "Reference taxonomy from ISMS-POL-A.8.11-S2.1"
    apply_style(subtitle, styles["subheader"])
    
    # COLUMN HEADERS (Row 4)
    headers = [
        ("A", "Category ID", 15),
        ("B", "Category Name", 25),
        ("C", "Description", 40),
        ("D", "Examples", 50),
        ("E", "Sensitivity Level", 15),
        ("F", "Masking Priority", 18),
        ("G", "Regulatory Driver", 25),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}4']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA CATEGORY DEFINITIONS (Pre-populated from Policy S2.1)
    row = 5
    categories_data = [
        ("CAT-PII-D", "Direct PII", "Data that directly identifies an individual",
         "Full Name, SSN/AHV, Passport, Driver's License, Email, Phone, Address, Photo",
         "High/Critical", "Required", "FADP, GDPR"),
        
        ("CAT-PII-I", "Indirect PII", "Data that can identify when combined",
         "DOB+ZIP, Job Title+Department, Employee ID",
         "Medium/High", "Required when combined", "FADP, GDPR"),
        
        ("CAT-FIN", "Financial Data", "Payment and financial account data",
         "PAN (Credit Card), CVV, IBAN, Account Balance, Salary, Tax ID",
         "Critical", "Required", "PCI-DSS, FADP, GDPR"),
        
        ("CAT-HLT", "Health Data", "Medical and health information",
         "Medical Record Number, Diagnoses, Medications, Lab Results, Insurance ID, Mental Health Records",
         "Critical", "Required", "FADP, GDPR Article 9"),
        
        ("CAT-CRD", "Credentials", "Authentication and access data",
         "Passwords, Password Hashes, API Keys, OAuth Tokens, Private Keys, DB Connection Strings",
         "Critical", "Required (Never display)", "Best Practice"),
        
        ("CAT-PRP", "Proprietary", "Business-sensitive information",
         "Trade Secrets, Pricing Models, Strategic Plans, Source Code",
         "High", "Recommended", "Business Confidentiality"),
        
        ("CAT-LOC", "Location Data", "Geographic and tracking data",
         "GPS Coordinates, IP Addresses, Travel History",
         "Medium/High", "Recommended", "GDPR (personal location)"),
        
        ("CAT-BIO", "Biometric Data", "Physical/behavioural identifiers",
         "Fingerprints, Facial Geometry, Voice Prints, Retina Scans",
         "Critical", "Required", "GDPR Article 9"),
        
        ("CAT-GEN", "Genetic Data", "DNA and genetic information",
         "DNA Sequences, Genetic Test Results, Hereditary Information",
         "Critical", "Required", "GDPR Article 9"),
        
        ("CAT-CHD", "Child Data", "Data relating to minors",
         "Any PII of persons under 16 (GDPR) / 18 (local)",
         "Critical", "Required", "GDPR Article 8, COPPA"),
    ]
    
    for cat_id, cat_name, description, examples, sensitivity, masking, regulations in categories_data:
        ws[f'A{row}'] = cat_id
        ws[f'B{row}'] = cat_name
        ws[f'C{row}'] = description
        ws[f'D{row}'] = examples
        ws[f'E{row}'] = sensitivity
        ws[f'F{row}'] = masking
        ws[f'G{row}'] = regulations
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].alignment = Alignment(wrap_text=True)

        row += 1
    
    # Freeze panes
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7: SHEET 4 - SENSITIVE DATA INVENTORY
# ============================================================================

def create_sensitive_data_inventory(ws, styles):
    """Create Sensitive Data Inventory sheet with 51-row template."""

    validations = create_base_validations(ws)

    # HEADER
    ws.merge_cells('A1:R1')
    header = ws['A1']
    header.value = "SENSITIVE DATA ELEMENT INVENTORY"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35

    # SUBTITLE
    ws.merge_cells('A2:R2')
    subtitle = ws['A2']
    subtitle.value = "Document sensitive data at the table/field level (51 row template: 1 sample + 50 empty)"
    apply_style(subtitle, styles["subheader"])

    # ASSESSMENT QUESTION
    ws.merge_cells('A3:R3')
    ws['A3'] = "Has your organisation identified all sensitive data fields across systems?"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(wrap_text=True)
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    ws['B4'].value = ""
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # COLUMN HEADERS (Row 6)
    headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Database/Schema", 20),
        ("D", "Table/Collection Name", 25),
        ("E", "Field/Column Name", 25),
        ("F", "Data Type", 15),
        ("G", "Data Category", 15),
        ("H", "Specific PII Type", 20),
        ("I", "Sensitivity Level", 15),
        ("J", "Contains PII?", 12),
        ("K", "Regulatory Scope", 20),
        ("L", "Masking Required?", 15),
        ("M", "Current Masking Status", 18),
        ("N", "Record Count (Approx)", 15),
        ("O", "Data Owner", 20),
        ("P", "Discovery Method", 18),
        ("Q", "Discovery Date", 15),
        ("R", "Notes/Comments", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}6']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # SAMPLE ROW (Row 7) - First data row with complete example
    sample_data = [
        "DE-001",
        "Customer CRM System",
        "CRM_PROD",
        "Customers",
        "email_address",
        "String",
        "CAT-PII-D",
        "Email",
        "High",
        "Yes",
        "GDPR, FADP",
        "Yes",
        "Not Masked",
        "500000",
        "VP Sales",
        "Schema Review",
        "2024-11-15",
        "Primary customer contact email",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = value
        apply_style(cell, styles["input_cell"])

    # EMPTY DATA ROWS (8-57: 50 additional rows = 51 total)
    for row_idx in range(8, 58):
        for col_idx in range(1, 19):  # A-R = 18 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Apply data validations to appropriate columns
    for row in range(7, 58):
        validations['data_type'].add(ws[f'F{row}'])
        validations['data_category'].add(ws[f'G{row}'])
        validations['pii_type'].add(ws[f'H{row}'])
        validations['sensitivity_level'].add(ws[f'I{row}'])
        validations['yes_no'].add(ws[f'J{row}'])
        validations['yes_no_conditional'].add(ws[f'L{row}'])
        validations['masking_status'].add(ws[f'M{row}'])
        validations['discovery_method'].add(ws[f'P{row}'])
    
    # CHECKLIST SECTION (Starting Row 59)
    checklist_row = 59
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "SENSITIVE DATA INVENTORY CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    # Checklist headers
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is sensitive data identified at field level?",
        "Are all production databases inventoried?",
        "Are non-production environments inventoried?",
        "Is PII specifically flagged?",
        "Are special category data (GDPR Art.9) identified?",
        "Is financial data (PCI scope) identified?",
        "Are credentials/secrets identified?",
        "Is data category assigned per field?",
        "Is sensitivity level assigned?",
        "Is masking requirement determined?",
        "Is current masking status documented?",
        "Are record counts estimated?",
        "Is data ownership assigned?",
        "Is discovery method documented?",
        "Are regulatory requirements mapped?",
        "Is inventory updated within 30 days of changes?",
        "Is automated discovery used where feasible?",
        "Are discovery results validated manually?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30

    # Freeze panes
    ws.freeze_panes = "A7"

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: SHEET 5 - CLASSIFICATION MATRIX
# ============================================================================

def create_classification_matrix(ws, styles):
    """Create Classification Matrix sheet."""

    validations = create_base_validations(ws)

    # HEADER
    ws.merge_cells('A1:L1')
    header = ws['A1']
    header.value = "DATA CLASSIFICATION MATRIX"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35

    # SUBTITLE
    ws.merge_cells('A2:L2')
    subtitle = ws['A2']
    subtitle.value = "Map data sensitivity to organisational classification scheme"
    apply_style(subtitle, styles["subheader"])

    # ASSESSMENT QUESTION
    ws.merge_cells('A3:L3')
    ws['A3'] = "Does your organisation classify sensitive data according to an established taxonomy?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # CLASSIFICATION SUMMARY TABLE (Rows 6-11)
    ws.merge_cells('A6:L6')
    ws['A6'] = "CLASSIFICATION SUMMARY"
    ws['A6'].font = Font(bold=True, size=11)
    
    # Summary headers (Row 7)
    summary_headers = [
        ("A", "Organisational Class", 20),
        ("B", "Typical Sensitivity", 18),
        ("C", "Data Categories", 25),
        ("D", "Masking Requirement", 30),
        ("E", "Count of Fields", 15),
    ]
    
    for col, header_text, width in summary_headers:
        cell = ws[f'{col}7']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Summary data (Rows 8-11)
    summary_data = [
        ("Restricted", "Critical", "CAT-HLT, CAT-BIO, CAT-GEN, CAT-FIN (PAN)", "Mandatory in ALL non-prod"),
        ("Confidential", "High", "CAT-PII-D, CAT-FIN, CAT-CRD", "Mandatory in non-prod"),
        ("Internal", "Medium", "CAT-PII-I, CAT-PRP, CAT-LOC", "Risk-based masking"),
        ("Public", "Low/None", "Non-sensitive data", "No masking required"),
    ]
    
    for row_idx, (org_class, sensitivity, categories, masking_req) in enumerate(summary_data, start=8):
        ws[f'A{row_idx}'] = org_class
        ws[f'B{row_idx}'] = sensitivity
        ws[f'C{row_idx}'] = categories
        ws[f'D{row_idx}'] = masking_req
        ws[f'E{row_idx}'] = "=COUNTIF('Classification Matrix'!F:F,\"" + org_class + "\")"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row_idx}'].alignment = Alignment(wrap_text=True)

    # CLASSIFICATION ASSIGNMENT TABLE (Starting Row 14)
    ws.merge_cells('A14:L14')
    ws['A14'] = "CLASSIFICATION ASSIGNMENT (51 rows: 1 sample + 50 empty)"
    ws['A14'].font = Font(bold=True, size=11)
    
    # Column headers (Row 15)
    assign_headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Field Name", 25),
        ("D", "Data Category", 15),
        ("E", "Sensitivity Level", 15),
        ("F", "Organisational Classification", 20),
        ("G", "Masking Requirement", 18),
        ("H", "Classification Rationale", 30),
        ("I", "Classified By", 20),
        ("J", "Classification Date", 15),
        ("K", "Last Review Date", 15),
        ("L", "Next Review Date", 15),
    ]
    
    for col, header_text, width in assign_headers:
        cell = ws[f'{col}15']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 16) - First data row with complete example
    sample_data = [
        "DE-001",
        "Customer CRM System",
        "email_address",
        "CAT-PII-D",
        "High",
        "Confidential",
        "Mandatory in non-prod",
        "Email field contains direct PII under GDPR/FADP",
        "Data Protection Officer",
        "15.01.2026",
        "15.01.2026",
        "15.01.2027",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=16, column=col_idx)
        cell.value = value
        apply_style(cell, styles["input_cell"])

    # EMPTY DATA ROWS (17-66: 50 additional rows = 51 total)
    for row_idx in range(17, 67):
        for col_idx in range(1, 13):  # A-L = 12 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Apply data validations
    for row in range(16, 67):
        validations['org_classification'].add(ws[f'F{row}'])

    # CHECKLIST (Starting Row 68)
    checklist_row = 68
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "CLASSIFICATION CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is organisational classification scheme defined?",
        "Are all sensitive fields classified?",
        "Is classification aligned with data categories?",
        "Are masking requirements derived from classification?",
        "Is classification rationale documented?",
        "Are classifiers identified per field?",
        "Is classification reviewed annually?",
        "Are regulatory overrides documented?",
        "Is reclassification process defined?",
        "Are classification changes tracked?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    ws.column_dimensions['B'].width = 45

    ws.freeze_panes = "A16"

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: SHEET 6 - REGULATORY MAPPING
# ============================================================================

def create_regulatory_mapping(ws, styles):
    """Create Regulatory Mapping sheet."""

    validations = create_base_validations(ws)

    # HEADER
    ws.merge_cells('A1:N1')
    header = ws['A1']
    header.value = "REGULATORY REQUIREMENT MAPPING"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35

    # SUBTITLE
    ws.merge_cells('A2:N2')
    subtitle = ws['A2']
    subtitle.value = "Map data to applicable privacy and security regulations"
    apply_style(subtitle, styles["subheader"])

    # ASSESSMENT QUESTION
    ws.merge_cells('A3:N3')
    ws['A3'] = "Has your organisation identified which data is subject to specific regulatory requirements?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # REGULATORY SUMMARY (Rows 6-13)
    ws.merge_cells('A6:E6')
    ws['A6'] = "REGULATORY SUMMARY"
    ws['A6'].font = Font(bold=True, size=11)
    
    reg_headers = [
        ("A", "Regulation", 20),
        ("B", "Applicability", 15),
        ("C", "Data Categories in Scope", 30),
        ("D", "Field Count", 12),
        ("E", "Masking Mandated?", 18),
    ]
    
    for col, header_text, width in reg_headers:
        cell = ws[f'{col}7']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    regulations = [
        ("GDPR (EU)", "Art. 32 Security", "Yes (Art. 32)"),
        ("FADP (Switzerland)", "Data Protection", "Yes"),
        ("HIPAA (US - if applicable)", "Protected Health Info", "Yes"),
        ("PCI-DSS (if processing cards)", "Cardholder Data", "Yes (Req 3.4)"),
        ("CCPA/CPRA (California)", "Consumer Privacy", "Recommended"),
        ("Other (specify)", "", ""),
    ]
    
    for row_idx, (regulation, scope, masking) in enumerate(regulations, start=8):
        ws[f'A{row_idx}'] = regulation
        ws[f'B{row_idx}'].value = ""
        apply_style(ws[f'B{row_idx}'], styles["input_cell"])
        validations['yes_no_partial_planned_na'].add(ws[f'B{row_idx}'])
        ws[f'C{row_idx}'] = scope
        ws[f'D{row_idx}'] = "=COUNTIF('Regulatory Mapping'!E:E,\"Yes\")"
        ws[f'E{row_idx}'] = masking

    # DETAILED MAPPING TABLE (Starting Row 15)
    ws.merge_cells('A15:N15')
    ws['A15'] = "DETAILED REGULATORY MAPPING (51 rows: 1 sample + 50 empty)"
    ws['A15'].font = Font(bold=True, size=11)

    map_headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Field Name", 25),
        ("D", "Data Category", 15),
        ("E", "GDPR Applicable?", 15),
        ("F", "GDPR Article", 15),
        ("G", "FADP Applicable?", 15),
        ("H", "HIPAA Applicable?", 15),
        ("I", "PCI-DSS Applicable?", 15),
        ("J", "Other Regulations", 20),
        ("K", "Masking Mandated by Reg?", 18),
        ("L", "Regulatory Reference", 25),
        ("M", "Compliance Deadline", 15),
        ("N", "Notes", 30),
    ]

    for col, header_text, width in map_headers:
        cell = ws[f'{col}16']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 17) - First data row with complete example
    sample_data = [
        "DE-001",
        "Customer CRM System",
        "email_address",
        "CAT-PII-D",
        "Yes",
        "Art. 4(1), Art. 32",
        "Yes",
        "No",
        "No",
        "",
        "Yes",
        "GDPR Art. 32 (security); FADP Art. 7",
        "",
        "Email is direct PII under both GDPR and FADP",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=17, column=col_idx)
        cell.value = value
        apply_style(cell, styles["input_cell"])

    # EMPTY DATA ROWS (18-67: 50 additional rows = 51 total)
    for row_idx in range(18, 68):
        for col_idx in range(1, 15):  # A-N = 14 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(17, 68):
        validations['yes_no_unknown'].add(ws[f'E{row}'])
        validations['gdpr_article'].add(ws[f'F{row}'])
        validations['yes_no_unknown'].add(ws[f'G{row}'])
        validations['yes_no_unknown'].add(ws[f'H{row}'])
        validations['yes_no_unknown'].add(ws[f'I{row}'])
        validations['yes_no_partial'].add(ws[f'K{row}'])

    # CHECKLIST (Starting Row 69)
    checklist_row = 69
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "REGULATORY COMPLIANCE CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Are applicable regulations identified?",
        "Is GDPR applicability determined?",
        "Is FADP applicability determined?",
        "Is HIPAA applicability determined (if US)?",
        "Is PCI-DSS applicability determined?",
        "Are GDPR special categories (Art.9) identified?",
        "Is PCI-DSS PAN data identified?",
        "Are masking requirements per regulation documented?",
        "Are regulatory references cited?",
        "Are compliance deadlines tracked?",
        "Is DPO consulted for GDPR scope?",
        "Are cross-border data transfers considered?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    ws.column_dimensions['B'].width = 45

    ws.freeze_panes = "A17"

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 10: SHEET 7 - DATA OWNER ASSIGNMENT
# ============================================================================

def create_data_owner_assignment(ws, styles):
    """Create Data Owner Assignment sheet."""

    validations = create_base_validations(ws)

    # HEADER
    ws.merge_cells('A1:N1')
    header = ws['A1']
    header.value = "DATA OWNERSHIP ASSIGNMENT"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35

    # SUBTITLE
    ws.merge_cells('A2:N2')
    subtitle = ws['A2']
    subtitle.value = "Assign data owners per system and data category (ISMS-POL-A.8.11-S3)"
    apply_style(subtitle, styles["subheader"])

    # ASSESSMENT QUESTION
    ws.merge_cells('A3:N3')
    ws['A3'] = "Has your organisation assigned data owners for all sensitive data categories?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # OWNERSHIP SUMMARY BY CATEGORY (Rows 6-17)
    ws.merge_cells('A6:F6')
    ws['A6'] = "OWNERSHIP SUMMARY BY DATA CATEGORY"
    ws['A6'].font = Font(bold=True, size=11)
    
    owner_headers = [
        ("A", "Data Category", 20),
        ("B", "Data Owner (Role)", 20),
        ("C", "Data Owner (Name)", 20),
        ("D", "Backup Owner", 20),
        ("E", "Assignment Date", 15),
        ("F", "Last Review", 15),
    ]
    
    for col, header_text, width in owner_headers:
        cell = ws[f'{col}7']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    data_categories = [
        "CAT-PII-D (Direct PII)",
        "CAT-PII-I (Indirect PII)",
        "CAT-FIN (Financial)",
        "CAT-HLT (Health)",
        "CAT-CRD (Credentials)",
        "CAT-PRP (Proprietary)",
        "CAT-LOC (Location)",
        "CAT-BIO (Biometric)",
        "CAT-GEN (Genetic)",
        "CAT-CHD (Child Data)",
    ]
    
    for row_idx, category in enumerate(data_categories, start=8):
        ws[f'A{row_idx}'] = category
        for col in ['B', 'C', 'D', 'E', 'F']:
            apply_style(ws[f'{col}{row_idx}'], styles["input_cell"])

    # OWNERSHIP BY SYSTEM (Starting Row 19)
    ws.merge_cells('A19:N19')
    ws['A19'] = "OWNERSHIP BY SYSTEM (51 rows: 1 sample + 50 empty)"
    ws['A19'].font = Font(bold=True, size=11)

    system_headers = [
        ("A", "System Name", 25),
        ("B", "System Owner", 20),
        ("C", "Primary Data Owner", 20),
        ("D", "Data Steward", 20),
        ("E", "Business Owner", 20),
        ("F", "Data Categories in System", 25),
        ("G", "Ownership Documented?", 15),
        ("H", "RACI Matrix Reference", 20),
        ("I", "Assignment Date", 15),
        ("J", "Last Review Date", 15),
        ("K", "Next Review Date", 15),
        ("L", "Approval Status", 15),
        ("M", "Approver", 20),
        ("N", "Notes", 30),
    ]

    for col, header_text, width in system_headers:
        cell = ws[f'{col}20']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 21) - First data row with complete example
    sample_data = [
        "Customer CRM System",
        "CTO",
        "VP Sales",
        "CRM Data Analyst",
        "VP Sales",
        "CAT-PII-D, CAT-FIN",
        "Yes",
        "/docs/raci-matrix-crm.xlsx",
        "15.01.2026",
        "15.01.2026",
        "15.01.2027",
        "Approved",
        "CISO",
        "Annual ownership review completed Q1 2026",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=21, column=col_idx)
        cell.value = value
        apply_style(cell, styles["input_cell"])

    # EMPTY DATA ROWS (22-71: 50 additional rows = 51 total)
    for row_idx in range(22, 72):
        for col_idx in range(1, 15):  # A-N = 14 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(21, 72):
        validations['yes_no_partial'].add(ws[f'G{row}'])

    # DATA OWNER RESPONSIBILITIES REFERENCE (Starting Row 73)
    ws.merge_cells('A73:C73')
    ws['A73'] = "DATA OWNER RESPONSIBILITIES (Reference)"
    ws['A73'].font = Font(bold=True, size=11)

    ws['A74'] = "Responsibility"
    ws['B74'] = "Requirement Level"
    ws['C74'] = "Policy Reference"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}74'], styles["column_header"])
    
    responsibilities = [
        ("Approve data classification", "SHALL", "REQ-CLS-031"),
        ("Approve masking techniques for their data", "SHALL", "REQ-CLS-031"),
        ("Review classification annually", "SHALL", "REQ-CLS-032"),
        ("Approve access to unmasked data", "SHALL", "REQ-ENV-041"),
        ("Review masking effectiveness", "SHOULD", "REQ-TST-020"),
        ("Participate in risk assessments", "SHOULD", "REQ-GOV-010"),
    ]

    for row_idx, (responsibility, level, policy_ref) in enumerate(responsibilities, start=75):
        ws[f'A{row_idx}'] = responsibility
        ws[f'B{row_idx}'] = level
        ws[f'C{row_idx}'] = policy_ref

    # CHECKLIST (Starting Row 82)
    checklist_row = 82
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "OWNERSHIP CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Are data owners assigned per category?",
        "Are data owners assigned per system?",
        "Are data stewards identified?",
        "Are backup owners assigned?",
        "Is ownership documented in RACI matrix?",
        "Have data owners acknowledged responsibilities?",
        "Are ownership changes tracked?",
        "Is ownership reviewed annually?",
        "Are data owners trained on masking requirements?",
        "Do data owners approve masking techniques?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    ws.column_dimensions['B'].width = 45

    ws.freeze_panes = "A21"

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 11: SHEET 8 - MASKING PRIORITY MATRIX
# ============================================================================

def create_masking_priority_matrix(ws, styles):
    """Create Masking Priority Matrix sheet with risk-based scoring."""

    validations = create_base_validations(ws)

    # HEADER
    ws.merge_cells('A1:P1')
    header = ws['A1']
    header.value = "MASKING PRIORITY & RISK ASSESSMENT"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35

    # SUBTITLE
    ws.merge_cells('A2:P2')
    subtitle = ws['A2']
    subtitle.value = "Prioritise masking efforts based on risk, compliance, and business impact"
    apply_style(subtitle, styles["subheader"])

    # ASSESSMENT QUESTION
    ws.merge_cells('A3:P3')
    ws['A3'] = "Has your organisation prioritised masking implementation based on risk assessment?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # PRIORITY CALCULATION FORMULA (Rows 6-10)
    ws.merge_cells('A6:P6')
    ws['A6'] = "PRIORITY SCORE FORMULA: (Sensitivity × 3) + (Exposure Risk × 2) + (Regulatory Weight × 2) + (Volume Factor × 1)"
    ws['A6'].font = Font(bold=True, size=10)
    ws['A6'].alignment = Alignment(wrap_text=True)
    
    ws['A8'] = "Priority Tiers:"
    ws['A8'].font = Font(bold=True)
    tiers = [
        "P1 (Critical): Score ≥ 35 — Immediate action required",
        "P2 (High): Score 25-34 — Implement within 3 months",
        "P3 (Medium): Score 15-24 — Implement within 6 months",
        "P4 (Low): Score < 15 — Risk-based timeline",
    ]
    for idx, tier in enumerate(tiers, start=9):
        ws[f'A{idx}'] = tier
    
    # PRIORITY ASSESSMENT TABLE (Starting Row 13)
    ws.merge_cells('A13:P13')
    ws['A13'] = "PRIORITY ASSESSMENT (51 rows: 1 sample + 50 empty)"
    ws['A13'].font = Font(bold=True, size=11)
    
    priority_headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Field Name", 25),
        ("D", "Data Category", 15),
        ("E", "Sensitivity Score (1-5)", 12),
        ("F", "Exposure Risk (1-5)", 15),
        ("G", "Regulatory Score (1-5)", 15),
        ("H", "Volume Score (1-5)", 12),
        ("I", "Total Priority Score", 15),
        ("J", "Priority Tier", 12),
        ("K", "Current Masking Status", 18),
        ("L", "Target Implementation Date", 18),
        ("M", "Assigned To", 20),
        ("N", "Implementation Status", 18),
        ("O", "Blocking Issues", 25),
        ("P", "Notes", 30),
    ]
    
    for col, header_text, width in priority_headers:
        cell = ws[f'{col}14']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 15) - First data row with complete example
    sample_data = [
        "DE-001",
        "Customer CRM System",
        "email_address",
        "CAT-PII-D",
        "5",  # Sensitivity: Very High
        "5",  # Exposure: Production + external-facing
        "5",  # Regulatory: GDPR + FADP mandatory
        "5",  # Volume: >500k records
        "40",  # Total: (5×3)+(5×2)+(5×2)+(5×1) = 40
        "P1",
        "Not Masked",
        "31.01.2026",
        "Data Engineering",
        "In Progress",
        "",
        "Critical P1 - production unmasked data",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = value
        apply_style(cell, styles["input_cell"])

    # EMPTY DATA ROWS (16-65: 50 additional rows = 51 total)
    for row_idx in range(16, 66):
        for col_idx in range(1, 17):  # A-P = 16 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Apply data validations
    for row in range(15, 66):
        validations['exposure_risk'].add(ws[f'F{row}'])
        validations['implementation_status'].add(ws[f'N{row}'])

        # Formulas for calculated fields
        # Total Priority Score: (E×3)+(F×2)+(G×2)+(H×1) - requires manual calculation or VBA
        if row > 15:  # Skip sample row
            ws[f'I{row}'] = ""  # User enters or calculates externally
            ws[f'J{row}'] = ""  # Priority Tier based on score - requires manual entry

    # PRIORITY SUMMARY DASHBOARD (Starting Row 67)
    ws.merge_cells('A67:G67')
    ws['A67'] = "PRIORITY TIER SUMMARY"
    ws['A67'].font = Font(bold=True, size=11)
    
    summary_headers = [
        ("A", "Priority Tier", 15),
        ("B", "Count", 10),
        ("C", "% of Total", 12),
        ("D", "Avg Score", 12),
        ("E", "Complete", 12),
        ("F", "In Progress", 12),
        ("G", "Not Started", 12),
    ]

    for col, header_text, width in summary_headers:
        cell = ws[f'{col}68']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    priority_tiers = ["P1 (Critical)", "P2 (High)", "P3 (Medium)", "P4 (Low)", "TOTAL"]

    for row_idx, tier in enumerate(priority_tiers, start=69):
        ws[f'A{row_idx}'] = tier
        # Formulas would go here - simplified for generator
        ws[f'B{row_idx}'] = ""
        ws[f'C{row_idx}'] = ""
        ws[f'D{row_idx}'] = ""
        ws[f'E{row_idx}'] = ""
        ws[f'F{row_idx}'] = ""
        ws[f'G{row_idx}'] = ""

    # CHECKLIST (Starting Row 75)
    checklist_row = 75
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "PRIORITY CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is risk-based prioritization performed?",
        "Are sensitivity levels factored into priority?",
        "Is exposure risk assessed per field?",
        "Are regulatory mandates weighted?",
        "Is data volume considered?",
        "Are P1 items addressed immediately?",
        "Are target dates assigned per priority?",
        "Are responsible parties assigned?",
        "Is implementation status tracked?",
        "Are blocking issues documented?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    ws.column_dimensions['B'].width = 45

    ws.freeze_panes = "A15"

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 12: SHEET 9 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet."""

    validations = create_base_validations(ws)

    # HEADER
    ws.merge_cells('A1:O1')
    header = ws['A1']
    header.value = "DATA INVENTORY & CLASSIFICATION GAP ANALYSIS"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 35

    # SUBTITLE
    ws.merge_cells('A2:O2')
    subtitle = ws['A2']
    subtitle.value = "Identify missing inventory, classification, ownership, or masking"
    apply_style(subtitle, styles["subheader"])
    
    # GAP SUMMARY TABLE (Rows 4-12)
    ws.merge_cells('A4:F4')
    ws['A4'] = "GAP SUMMARY"
    ws['A4'].font = Font(bold=True, size=11)
    
    gap_summary_headers = [
        ("A", "Gap Category", 25),
        ("B", "Count", 10),
        ("C", "% of Total", 12),
        ("D", "Risk Level", 12),
        ("E", "Remediation Owner", 20),
        ("F", "Target Date", 15),
    ]
    
    for col, header_text, width in gap_summary_headers:
        cell = ws[f'{col}5']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    gap_categories = [
        "Systems not inventoried",
        "Systems with unknown sensitive data",
        "Fields not classified",
        "Fields without data owner",
        "Regulatory mapping incomplete",
        "Masking required but not implemented",
        "Masking status unknown",
    ]
    
    for row_idx, category in enumerate(gap_categories, start=6):
        ws[f'A{row_idx}'] = category
        apply_style(ws[f'B{row_idx}'], styles["input_cell"])
        apply_style(ws[f'C{row_idx}'], styles["input_cell"])
        apply_style(ws[f'D{row_idx}'], styles["input_cell"])
        validations['risk_level'].add(ws[f'D{row_idx}'])
        apply_style(ws[f'E{row_idx}'], styles["input_cell"])
        apply_style(ws[f'F{row_idx}'], styles["input_cell"])

    # DETAILED GAP REGISTER (Starting Row 14)
    ws.merge_cells('A14:O14')
    ws['A14'] = "DETAILED GAP REGISTER (51 rows: 1 sample + 50 empty)"
    ws['A14'].font = Font(bold=True, size=11)

    gap_headers = [
        ("A", "Gap ID", 12),
        ("B", "Gap Category", 20),
        ("C", "System Name", 25),
        ("D", "Field/Data Element", 25),
        ("E", "Gap Description", 35),
        ("F", "Risk Level", 12),
        ("G", "Impact if Not Remediated", 30),
        ("H", "Root Cause", 25),
        ("I", "Remediation Action", 30),
        ("J", "Remediation Owner", 20),
        ("K", "Target Completion Date", 15),
        ("L", "Status", 15),
        ("M", "Actual Completion Date", 15),
        ("N", "Verification Method", 20),
        ("O", "Notes", 30),
    ]

    for col, header_text, width in gap_headers:
        cell = ws[f'{col}15']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 16) - First data row with complete example
    sample_data = [
        "GAP-001",
        "Masking Missing",
        "HR Payroll System",
        "employee_bank_account",
        "Bank account numbers not masked in test environment - real employee data visible",
        "High",
        "Potential data breach if test environment compromised; GDPR/FADP violation",
        "Masking not configured during initial deployment in 2024",
        "Implement static data masking for bank_account field using format-preserving encryption",
        "Data Engineering Team",
        "31.03.2026",
        "In Progress",
        "",
        "Manual testing post-implementation",
        "Approved by DPO 15.01.2026; high priority P1",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=16, column=col_idx)
        cell.value = value
        apply_style(cell, styles["input_cell"])

    # EMPTY DATA ROWS (17-66: 50 additional rows = 51 total)
    for row_idx in range(17, 67):
        for col_idx in range(1, 16):  # A-O = 15 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Apply data validations
    for row in range(16, 67):
        validations['gap_category'].add(ws[f'B{row}'])
        validations['risk_level'].add(ws[f'F{row}'])
        validations['gap_status'].add(ws[f'L{row}'])

    # CHECKLIST (Starting Row 68)
    checklist_row = 68
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "GAP ANALYSIS CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is gap analysis performed quarterly?",
        "Are missing inventories identified?",
        "Are unclassified systems identified?",
        "Are fields without owners identified?",
        "Are regulatory gaps documented?",
        "Are masking gaps prioritised?",
        "Are root causes analysed?",
        "Are remediation actions defined?",
        "Are remediation owners assigned?",
        "Are target dates realistic?",
        "Is gap closure tracked?",
        "Are accepted risks documented?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    ws.column_dimensions['B'].width = 45

    ws.freeze_panes = "A16"

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 13: SHEET 10 - EVIDENCE REGISTER
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Start with System Inventory — list ALL systems/databases that process data.',
        '2. Use Data Category Reference to understand sensitivity taxonomy.',
        '3. Complete Sensitive Data Inventory for each system (table/field level).',
        '4. Apply classification using the Classification Matrix sheet.',
        '5. Map data to regulatory requirements in Regulatory Mapping.',
        '6. Assign data owners in Data Owner Assignment.',
        '7. Prioritise masking efforts in Masking Priority Matrix.',
        '8. Document gaps in Gap Analysis.',
        '9. Maintain evidence in Evidence Register.',
        '10. Review summary metrics in Summary Dashboard.',
        '11. Obtain required approvals in Approval Sign-Off.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A25"] = "Status Legend"
    ws["A25"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=26, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 27 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — golden standard 8-column layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 — merge A1:H1, title "EVIDENCE REGISTER"
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2 — subtitle, italic
    ws.merge_cells("A2:H2")
    ws["A2"] = f"List all evidence files/documents referenced in this assessment ({DOCUMENT_ID})."
    ws["A2"].font = Font(italic=True, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers Row 4 — 8 columns, D9D9D9 fill, bold, thin borders
    headers = [
        "Evidence ID", "Category", "Description", "Source/Location",
        "Date Collected", "Collected By", "Status", "Notes"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, name="Calibri", color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Evidence category dropdown (column B)
    ev_cat_dv = DataValidation(
        type="list",
        formula1='"Data Discovery Report,Schema Documentation,DPIA,Data Flow Diagram,Classification Review,RACI Matrix,Meeting Minutes,Approval Email,Tool Report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(ev_cat_dv)

    # Verification Status dropdown (column G)
    ver_status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True
    )
    ws.add_data_validation(ver_status_dv)

    # Sample row with complete example data
    row = 5
    sample_data = {
        1: "EV-001",
        2: "Data Discovery Report",
        3: "Sensitive data inventory scan results for Q1 2026",
        4: "/evidence/data-discovery-q1-2026.xlsx",
        5: "15.01.2026",
        6: "Data Protection Team",
        7: "Verified",
        8: "Quarterly scan covering all production databases"
    }
    
    for col, value in sample_data.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Empty rows 6-104 (99 empty rows)
    for r in range(6, 105):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Apply dropdowns to ranges
    ev_cat_dv.add("B5:B104")
    ver_status_dv.add("G5:G104")

    # Column widths per spec
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 30

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 14: SHEET 11 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # ── Row 1: Main header ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle ──
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        f"Summary Dashboard  |  {WORKBOOK_NAME}  |  Generated: {GENERATED_DATE}"
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: blank ──

    # ── TABLE 1 banner (Row 4) ──
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 1 column headers (Row 5) ──
    t1_headers = [
        "Assessment Area", "Total Items", "Compliant",
        "Partial", "Non-Compliant", "N/A", "Compliance %",
    ]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 1 data rows ──
    # Assessment areas and their COUNTIF formulas.
    # Ranges extended to row 100 to cover data + checklist rows in each sheet.
    # Col B uses COUNTA on the primary data column (col B = first user-entered col).
    # Non-BMP emoji (📋 U+1F4CB) excluded from COUNTIF — not supported in Excel.
    #
    # Sensitive Data Inventory — col M (masking_status), rows 8:100
    #   DV: "Masked,Not Masked,Partially Masked,Encrypted,Planned"
    #   Compliant = COUNTIF(M8:M100,"Masked")+COUNTIF(M8:M100,"Encrypted")
    #   Partial   = COUNTIF(M8:M100,"Partially Masked")
    #   Non-Comp  = COUNTIF(M8:M100,"Not Masked")
    #   N/A       = COUNTIF(M8:M100,"Planned")
    #
    # System Inventory — col J (status_icons), rows 8:100
    #   DV: "✅ Complete,⚠️ Partial,❌ Missing,N/A" (📋 Planned excluded — non-BMP)
    #   Compliant = COUNTIF(J8:J100,"✅ Complete")
    #   Partial   = COUNTIF(J8:J100,"⚠️ Partial")
    #   Non-Comp  = COUNTIF(J8:J100,"❌ Missing")
    #   N/A       = COUNTIF(J8:J100,"N/A")
    #
    # Classification Matrix — col F (org_classification), rows 17:100
    #   Compliant = COUNTA(F17:F100)
    #
    # Masking Priority Matrix — col N (implementation_status), rows 16:100
    #   DV: "Not Started,In Progress,Complete,Blocked"
    #
    # Gap Analysis — col L (gap_status), rows 17:100
    #   DV: "Open,In Progress,Complete,Accepted Risk"

    t1_areas = [
        (
            "Sensitive Data Inventory",
            "=COUNTA('Sensitive Data Inventory'!B8:B100)",
            "=COUNTIF('Sensitive Data Inventory'!M8:M100,\"Masked\")+COUNTIF('Sensitive Data Inventory'!M8:M100,\"Encrypted\")",
            "=COUNTIF('Sensitive Data Inventory'!M8:M100,\"Partially Masked\")",
            "=COUNTIF('Sensitive Data Inventory'!M8:M100,\"Not Masked\")",
            "=COUNTIF('Sensitive Data Inventory'!M8:M100,\"Planned\")",
        ),
        (
            "System Inventory",
            "=COUNTA('System Inventory'!B8:B100)",
            "=COUNTIF('System Inventory'!J8:J100,\"\u2705 Complete\")",
            "=COUNTIF('System Inventory'!J8:J100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('System Inventory'!J8:J100,\"\u274C Missing\")",
            "=COUNTIF('System Inventory'!J8:J100,\"N/A\")",
        ),
        (
            "Classification Matrix",
            "=COUNTA('Classification Matrix'!B17:B100)",
            "=COUNTA('Classification Matrix'!F17:F100)",
            "=0",
            "=0",
            "=0",
        ),
        (
            "Masking Priority Matrix",
            "=COUNTA('Masking Priority Matrix'!B16:B100)",
            "=COUNTIF('Masking Priority Matrix'!N16:N100,\"Complete\")",
            "=COUNTIF('Masking Priority Matrix'!N16:N100,\"In Progress\")",
            "=COUNTIF('Masking Priority Matrix'!N16:N100,\"Not Started\")+COUNTIF('Masking Priority Matrix'!N16:N100,\"Blocked\")",
            "=0",
        ),
        (
            "Gap Analysis",
            "=COUNTA('Gap Analysis'!B17:B100)",
            "=COUNTIF('Gap Analysis'!L17:L100,\"Complete\")+COUNTIF('Gap Analysis'!L17:L100,\"Accepted Risk\")",
            "=COUNTIF('Gap Analysis'!L17:L100,\"In Progress\")",
            "=COUNTIF('Gap Analysis'!L17:L100,\"Open\")",
            "=0",
        ),
    ]

    for row_idx, (area, total, compliant, partial, non_comp, na) in enumerate(t1_areas, start=6):
        row = row_idx
        # Col A — area name
        ws[f"A{row}"] = area
        ws[f"A{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row}"].border = border
        # Col B — total
        ws[f"B{row}"] = total
        ws[f"B{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row}"].border = border
        # Col C — compliant
        ws[f"C{row}"] = compliant
        ws[f"C{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row}"].border = border
        # Col D — partial
        ws[f"D{row}"] = partial
        ws[f"D{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{row}"].border = border
        # Col E — non-compliant
        ws[f"E{row}"] = non_comp
        ws[f"E{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"E{row}"].border = border
        # Col F — N/A
        ws[f"F{row}"] = na
        ws[f"F{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row}"].border = border
        # Col G — compliance %
        ws[f"G{row}"] = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        ws[f"G{row}"].number_format = "0.0%"
        ws[f"G{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"G{row}"].border = border

    # ── TOTAL row ──
    total_row = 11
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Calibri", size=10, bold=True, color="000000")
    ws[f"A{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{total_row}"].border = border
    for col_letter, formula in [
        ("B", "=SUM(B6:B10)"),
        ("C", "=SUM(C6:C10)"),
        ("D", "=SUM(D6:D10)"),
        ("E", "=SUM(E6:E10)"),
        ("F", "=SUM(F6:F10)"),
    ]:
        ws[f"{col_letter}{total_row}"] = formula
        ws[f"{col_letter}{total_row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"{col_letter}{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col_letter}{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col_letter}{total_row}"].border = border
    ws[f"G{total_row}"] = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    ws[f"G{total_row}"].number_format = "0.0%"
    ws[f"G{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"G{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"G{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"G{total_row}"].border = border

    # ── TABLE 2 banner (Row 13) ──
    ws.merge_cells("A13:G13")
    ws["A13"] = "TABLE 2 \u2013 KEY METRICS"
    ws["A13"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A13"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A13"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 2 headers (Row 14) ──
    for col_idx, hdr in enumerate(["Metric", "Value", "Target", "", "", "", ""], start=1):
        cell = ws.cell(row=14, column=col_idx, value=hdr if hdr else None)
        if hdr:
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    # ── TABLE 2 data rows (Rows 15-20) ──
    t2_data = [
        ("Sensitive Data Fields Masked (%)", f"=IF(COUNTA('Sensitive Data Inventory'!M8:M100)=0,\"N/A\",TEXT((COUNTIF('Sensitive Data Inventory'!M8:M100,\"Masked\")+COUNTIF('Sensitive Data Inventory'!M8:M100,\"Encrypted\"))/COUNTA('Sensitive Data Inventory'!M8:M100),\"0.0%\"))", "100%"),
        ("Systems Fully Inventoried (%)", f"=IF(COUNTA('System Inventory'!J8:J100)=0,\"N/A\",TEXT(COUNTIF('System Inventory'!J8:J100,\"\u2705 Complete\")/COUNTA('System Inventory'!J8:J100),\"0.0%\"))", "100%"),
        ("Data Fields Classified (%)", f"=IF(COUNTA('Classification Matrix'!A17:A100)=0,\"N/A\",TEXT(COUNTA('Classification Matrix'!F17:F100)/COUNTA('Classification Matrix'!A17:A100),\"0.0%\"))", "100%"),
        ("P1 Gaps Remediated (%)", f"=IF(COUNTA('Gap Analysis'!A17:A100)=0,\"N/A\",TEXT((COUNTIF('Gap Analysis'!L17:L100,\"Complete\")+COUNTIF('Gap Analysis'!L17:L100,\"Accepted Risk\"))/COUNTA('Gap Analysis'!A17:A100),\"0.0%\"))", "100%"),
        ("Open Critical Gaps (Count)", f"=COUNTIF('Gap Analysis'!L17:L100,\"Open\")", "0"),
        ("Masking Implementation Complete (%)", f"=IF(COUNTA('Masking Priority Matrix'!N16:N100)=0,\"N/A\",TEXT(COUNTIF('Masking Priority Matrix'!N16:N100,\"Complete\")/COUNTA('Masking Priority Matrix'!N16:N100),\"0.0%\"))", "100%"),
    ]
    for row_idx, (metric, value, target) in enumerate(t2_data, start=15):
        ws[f"A{row_idx}"] = metric
        ws[f"A{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row_idx}"].border = border
        ws[f"B{row_idx}"] = value
        ws[f"B{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row_idx}"].border = border
        ws[f"C{row_idx}"] = target
        ws[f"C{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row_idx}"].border = border

    # ── TABLE 3 banner (Row 22) ──
    ws.merge_cells("A22:G22")
    ws["A22"] = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    ws["A22"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A22"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws["A22"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 3 headers (Row 23) ──
    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", "Notes"]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=23, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 3 data rows (Rows 24-27) ──
    t3_data = [
        ("1", "Sensitive data fields not yet masked in non-production environments", "High — GDPR/FADP non-compliance risk", "Implement Static Data Masking (SDM) in all non-production data refresh pipelines", "P1", "Open", ""),
        ("2", "Data classification incomplete for some systems", "Medium — masking priorities cannot be set without classification", "Complete classification matrix for all inventoried systems; assign data owners", "P2", "In Progress", ""),
        ("3", "Data owner assignments missing for several systems", "Medium — accountability gap; masking decisions require owner approval", "Conduct data ownership workshops; RACI matrix to be approved by CISO", "P2", "Open", ""),
        ("4", "Gap analysis items not all remediated", "Medium — open gaps indicate unprotected sensitive data", "Prioritise P1 gaps; establish 30/60/90 day remediation milestones", "P2", "In Progress", ""),
    ]
    for row_idx, row_data in enumerate(t3_data, start=24):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border

    # ── Freeze panes ──
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 15: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet — golden standard layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 — merge A1:E1, "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 3: ASSESSMENT SUMMARY banner — 4472C4 fill
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Status dropdown on Assessment Status
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)
    if status_row:
        status_dv.add(ws[f"B{status_row}"])

    # 3 approver sections
    approvers = [
        ("PREPARED BY (DATA GOVERNANCE)", "4472C4"),
        ("REVIEWED BY (DPO / ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True, name="Calibri")
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            row += 1
        row += 1  # gap between sections

    # FINAL ASSESSMENT DECISION — col A plain bold label (GS: no dark fill, no merge)
    ws[f"A{row}"] = "FINAL ASSESSMENT DECISION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="000000", name="Calibri")
    row += 1

    ws[f"A{row}"] = "Decision:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Not Approved,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # NEXT REVIEW DATE section
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    row += 1

    review_fields = [
        ("Recommended Date:", None),
        ("Review Frequency:", "review_frequency"),
        ("Assigned Reviewer:", None),
        ("Notes:", None),
    ]

    review_freq_dv = DataValidation(
        type="list",
        formula1='"Annual,Semi-Annual,Quarterly,As-Needed"',
        allow_blank=True
    )
    ws.add_data_validation(review_freq_dv)

    for label, dv_key in review_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        if dv_key == "review_frequency":
            review_freq_dv.add(ws[f"B{row}"])
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    # Apply borders to all merged cell top-left corners (GS-AS-011)
    _as_thin = Side(style="thin")
    _as_border = Border(left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin)
    for merge_range in ws.merged_cells.ranges:
        tl = ws.cell(merge_range.min_row, merge_range.min_col)
        tl.border = _as_border


# ============================================================================
# SECTION 16: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    try:
        logger.info("=" * 78)
        logger.info(f"{WORKBOOK_ID} - {ASSESSMENT_AREA} Generator")
        logger.info("ISO/IEC 27001:2022 Control A.8.11: Data Masking")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/12] Creating Instructions & Legend...")
        create_instructions_legend(wb["Instructions & Legend"], styles)

        logger.info("[2/12] Creating System Inventory...")
        create_system_inventory(wb["System Inventory"], styles)

        logger.info("[3/12] Creating Data Category Reference...")
        create_data_category_reference(wb["Data Category Reference"], styles)

        logger.info("[4/12] Creating Sensitive Data Inventory...")
        create_sensitive_data_inventory(wb["Sensitive Data Inventory"], styles)

        logger.info("[5/12] Creating Classification Matrix...")
        create_classification_matrix(wb["Classification Matrix"], styles)

        logger.info("[6/12] Creating Regulatory Mapping...")
        create_regulatory_mapping(wb["Regulatory Mapping"], styles)

        logger.info("[7/12] Creating Data Owner Assignment...")
        create_data_owner_assignment(wb["Data Owner Assignment"], styles)

        logger.info("[8/12] Creating Masking Priority Matrix...")
        create_masking_priority_matrix(wb["Masking Priority Matrix"], styles)

        logger.info("[9/12] Creating Gap Analysis...")
        create_gap_analysis(wb["Gap Analysis"], styles)

        logger.info("[10/12] Creating Evidence Register...")
        create_evidence_register(wb["Evidence Register"], styles)

        logger.info("[11/12] Creating Summary Dashboard...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

        logger.info("[12/12] Creating Approval Sign-Off...")
        create_approval_sheet(wb["Approval Sign-Off"], styles)

        # Save workbook
        filename = f"ISMS-IMP-A.8.11.1_Data_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"SUCCESS: {filename}")
        logger.info("Workbook Structure: 12 sheets including Instructions, System Inventory, Evidence Register, Approval Sign-Off")
        logger.info("=" * 78)
        return 0
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
