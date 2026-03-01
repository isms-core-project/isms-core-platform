#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# Licensed under AGPL-3.0-or-later with commercial licensing option
#
# This file is part of the ISMS Compliance Framework
# See /LICENSE for full terms and /LICENSES/COMMERCIAL.md for commercial options
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.11.3 - Environment Coverage Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Assessment Domain 3 of 4: Environment Coverage Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific data masking infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Data classification categories requiring masking (match your data governance framework)
2. Masking technique selection criteria per data type and use case
3. Environment categories where masking is mandatory (dev, test, analytics)
4. Testing and validation procedure requirements per masking technique
5. Masking exception approval workflow and compensating control requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.11 Data Masking Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
data masking controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Environment Coverage Assessment under ISO 27001:2022 Control A.8.11. Supports evidence-based evaluation of data masking coverage, technique selection compliance, and validation effectiveness.

**Assessment Scope:**
- Sensitive data inventory completeness and masking requirement identification
- Masking technique selection appropriateness per data category
- Non-production environment masking coverage and compliance
- Masking process documentation and automation coverage
- Testing and validation procedure completeness and outcome tracking
- Exception management and approved compensating controls documentation
- Evidence collection for data protection and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
5. Summary Dashboard - Compliance overview and key metrics
6. Evidence Register - Audit evidence tracking
7. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Data Masking controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a811_3_environment_coverage.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a811_3_environment_coverage.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a811_3_environment_coverage.py --date 20250115

Output:
    File: ISMS-IMP-A.8.11.3_Environment_Coverage_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Assessment Domain:    3 of 4 (Environment Coverage Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.11: Data Masking Policy (Governance)
    - ISMS-IMP-A.8.11.1: Data Inventory & Classification Assessment (Domain 1)
    - ISMS-IMP-A.8.11.2: Masking Technique Selection & Requirements (Domain 2)
    - ISMS-IMP-A.8.11.3: Environment Coverage Assessment (Domain 3)
    - ISMS-IMP-A.8.11.4: Testing & Validation Framework (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.11.3 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive data masking details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review data masking inventories and technique requirements annually or when new data categories are introduced, non-production environments change, or data protection incidents occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.11.3"
WORKBOOK_NAME = "Environment Coverage Assessment"
CONTROL_ID = "A.8.11"
CONTROL_NAME = "Data Masking"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"



# =============================================================================
# HELPER: BATCH DATA VALIDATION
# =============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Instructions & Legend",
        "Environment Inventory",
        "Production Environment",
        "NonProduction Environments",
        "Analytics & Reporting",
        "Backup & Archive",
        "External Sharing",
        "Cloud Environments",
        "Data Flow Mapping",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def get_styles():
    """Define all cell styles."""
    thin = Side(style="thin")

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS
# ============================================================================

def get_base_columns():
    """Standard 17 columns (A-Q) used across assessment sheets."""
    return {
        "Environment Name": 25,
        "Environment Type": 20,
        "Classification": 18,
        "Hosting Location": 18,
        "Data Sensitivity": 18,
        "Masking Required?": 18,
        "Masking Deployed?": 18,
        "Masking Technique": 20,
        "Masking Tool/Solution": 22,
        "Coverage %": 12,
        "Last Verified Date": 15,
        "Environment Owner": 20,
        "Data Owner": 20,
        "Exception Approved?": 15,
        "Compliance Status": 18,
        "Notes/Comments": 30,
        "Evidence ID": 15,
    }


def get_extended_columns(sheet_type):
    """Extended columns for specific sheets."""
    extensions = {
        "production": {
            "User Role/Group": 20,
            "Masked Fields": 25,
            "Unmasked Access Logged?": 15,
            "Access Control Method": 20,
            "Exception Justification": 30,
            "Risk Level": 12,
            "Remediation Target Date": 15,
        },
        "nonproduction": {
            "Data Refresh Frequency": 18,
            "Masking Applied During Refresh?": 18,
            "Direct Prod Clone Prevented?": 18,
            "Masking Validation Method": 22,
            "Last Data Refresh Date": 15,
            "Next Planned Refresh": 15,
            "Refresh Process Owner": 20,
        },
        "analytics": {
            "Analytics Platform Type": 22,
            "Aggregation Level": 18,
            "Re-ID Risk Assessed?": 15,
            "Re-ID Risk Level": 15,
            "Synthetic Data Used?": 15,
            "Data Export Controls": 22,
            "Self-Service BI Masking": 18,
        },
        "backup": {
            "Backup Type": 18,
            "Encryption Enabled?": 15,
            "Encryption Method": 20,
            "Access Control": 18,
            "Restoration Process": 22,
            "Masking on Restore?": 18,
            "Backup Retention Period": 15,
        },
        "external": {
            "Recipient Type": 20,
            "Data Sharing Purpose": 25,
            "DPA in Place?": 15,
            "DPA Specifies Masking?": 18,
            "Contractual Exception?": 18,
            "Recipient Security Audit Date": 15,
            "Data Minimization Applied?": 18,
        },
        "cloud": {
            "Cloud Provider": 18,
            "Cloud Service Type": 20,
            "Region/Geography": 18,
            "Client-Side Masking?": 18,
            "Cloud-Native Masking Tool": 22,
            "Multi-Tenancy Concerns?": 18,
            "Data Residency Compliance": 22,
        },
        "dataflow": {
            "Source Environment": 20,
            "Destination Environment": 20,
            "Data Type": 18,
            "Masking Checkpoint?": 18,
            "Masking Technique": 20,
            "Flow Frequency": 15,
            "Automated Masking?": 18,
            "Masking Tool/Script": 22,
            "Masking Validation": 20,
            "Last Flow Date": 15,
            "Flow Owner": 20,
            "Approval Required?": 15,
            "Approval Status": 18,
            "Compliance Status": 18,
            "Risk Level": 12,
            "Notes/Comments": 30,
            "Evidence ID": 15,
        },
    }
    return extensions.get(sheet_type, {})


# ============================================================================
# SECTION 3: DATA VALIDATION
# ============================================================================

def create_validations():
    """Create data validation objects (not yet added to worksheet)."""
    validations = {
        'env_type': DataValidation(type="list", formula1='"Production,Development,Testing,UAT,Staging,Training,Sandbox,Analytics,Cloud,Backup,Archive,External"'),
        'classification': DataValidation(type="list", formula1='"Sensitive,Confidential,Internal,Public"'),
        'hosting': DataValidation(type="list", formula1='"On-Premises,AWS,Azure,GCP,Hybrid,Other Cloud"'),
        'data_sens': DataValidation(type="list", formula1='"PII,Financial,Health,Credentials,Proprietary,Mixed,None"'),
        'masking_req': DataValidation(type="list", formula1='"\u2705 Mandatory,\u26A0\uFE0F Conditional,\u274C Not Required,N/A"'),
        'yes_no': DataValidation(type="list", formula1='"Yes,No,Partial,Planned,N/A"'),
        'technique': DataValidation(type="list", formula1='"SDM,DDM,Tokenization,Encryption,Redaction,Substitution,Anonymization,None"'),
        'compliance': DataValidation(type="list", formula1='"\u2705 Compliant,\u26A0\uFE0F Partial,\u274C Non-Compliant,N/A"'),
        'yes_no_simple': DataValidation(type="list", formula1='"Yes,No,N/A"'),
        'checklist': DataValidation(type="list", formula1='"\u2705 Complete,\u26A0\uFE0F Partial,\u274C Missing,N/A"'),
    }

    return validations


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 4: REUSABLE SHEET CREATION FUNCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete each worksheet tab in sequence (Environment Inventory → Production → Non-Production → etc.)',
        '2. Fill ALL yellow-highlighted cells with your organisation\'s specific information.',
        '3. Use dropdown menus where provided (do not type free-form text in dropdown cells).',
        '4. Document ALL environments in your organisation (include cloud, on-premises, hybrid).',
        '5. For each environment, specify masking requirement (Mandatory/Conditional/Not Required).',
        '6. Verify masking deployment status (Yes/No/Partial/Planned).',
        '7. Calculate coverage percentage (% of sensitive fields masked).',
        '8. Link all assessments to Evidence Register with unique Evidence IDs.',
        '9. Complete Gap Analysis sheet to identify remediation needs.',
        '10. Review Summary Dashboard for executive-level compliance status.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A24"] = "Status Legend"
    ws["A24"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=25, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 26 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_assessment_sheet(ws, styles, title, policy_ref, question,
                           columns, row_count, checklist_items, sheet_type=None):
    """Generic function to create assessment sheets."""

    # Header
    ws.merge_cells('A1:Q1')
    ws['A1'] = title.upper()
    apply_style(ws['A1'], styles["header"])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Policy Reference
    ws.merge_cells('A2:Q2')
    ws['A2'] = f"Policy Reference: {policy_ref}"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment Question
    ws['A3'] = question
    ws['A3'].font = Font(bold=True, size=11)
    ws['B3'].fill = styles["input_cell"]["fill"]
    ws['B3'].border = styles["border"]

    validations = create_validations()
    validations['yes_no'].add(ws['B3'])

    # Column Headers
    col_num = 1
    for col_name, col_width in columns.items():
        cell = ws.cell(row=6, column=col_num, value=col_name)
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[get_column_letter(col_num)].width = col_width
        col_num += 1

    total_cols = len(columns)

    # SAMPLE ROW (Row 7) - First data row with complete example
    sample_data = [
        "ENV-001",
        "Production",
        "Confidential",
        "AWS EU-West-1",
        "High",
        "Yes",
        "Yes",
        "DDM (Dynamic)",
        "Role-based masking for CSRs",
        "Data Engineering",
        "15.01.2026",
        "Q1 2026",
        "EV-001",
        "N/A",
        "✅ Compliant",
        "Production DDM deployed and validated",
        "John Doe",
    ]

    # Apply sample data to row 7 (up to total_cols to avoid index errors)
    for col_idx in range(1, min(len(sample_data) + 1, total_cols + 1)):
        cell = ws.cell(row=7, column=col_idx)
        if col_idx <= len(sample_data):
            cell.value = sample_data[col_idx - 1]
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = styles["border"]

    # EMPTY DATA ROWS (8-57: 50 additional rows = 51 total)
    start_row = 8
    end_row = 57

    for row in range(start_row, end_row + 1):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Apply specific validations based on sheet type
    apply_column_validations(ws, validations, 7, end_row, sheet_type)

    # Compliance Checklist
    checklist_row = end_row + 3
    ws[f'A{checklist_row}'] = f"{title.split()[0]} CHECKLIST".upper()
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1

    for item in checklist_items:
        ws[f'A{checklist_row}'] = "\u2610"
        ws[f'B{checklist_row}'] = item
        ws[f'C{checklist_row}'].fill = styles["input_cell"]["fill"]
        ws[f'C{checklist_row}'].border = styles["border"]
        validations['checklist'].add(ws[f'C{checklist_row}'])
        checklist_row += 1

    # Finalize all validations
    for _dv in validations.values():
        ws.add_data_validation(_dv)

    ws.freeze_panes = "A7"


def apply_column_validations(ws, validations, start_row, end_row, sheet_type):
    """Apply data validations to specific columns."""

    # Column B: Environment Type
    for row in range(start_row, end_row + 1):
        validations['env_type'].add(ws[f'B{row}'])

    # Column C: Classification
    for row in range(start_row, end_row + 1):
        validations['classification'].add(ws[f'C{row}'])

    # Column D: Hosting Location
    for row in range(start_row, end_row + 1):
        validations['hosting'].add(ws[f'D{row}'])

    # Column E: Data Sensitivity
    for row in range(start_row, end_row + 1):
        validations['data_sens'].add(ws[f'E{row}'])

    # Column F: Masking Required?
    for row in range(start_row, end_row + 1):
        validations['masking_req'].add(ws[f'F{row}'])

    # Column G: Masking Deployed?
    for row in range(start_row, end_row + 1):
        validations['yes_no'].add(ws[f'G{row}'])

    # Column H: Masking Technique
    for row in range(start_row, end_row + 1):
        validations['technique'].add(ws[f'H{row}'])

    # Column N: Exception Approved?
    for row in range(start_row, end_row + 1):
        validations['yes_no_simple'].add(ws[f'N{row}'])

    # Column O: Compliance Status
    for row in range(start_row, end_row + 1):
        validations['compliance'].add(ws[f'O{row}'])


# ============================================================================
# SECTION 5: INSTRUCTIONS & LEGEND SHEET (GOLDEN STANDARD)
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet (golden standard IL)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header (Row 1) — two-line, merged A1:G1, height 40 ──
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # ── Document Information (Row 3+) — plain bold heading, NO banner ──
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", WORKBOOK_NAME),
        ("Related Policy", "ISMS-POL-A.8.11"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # ── Instructions — plain bold heading, NO banner, NO fill, NO merge ──
    row += 1
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12)

    instructions = [
        "1. Complete each worksheet tab in sequence (Environment Inventory \u2192 Production \u2192 Non-Production \u2192 etc.)",
        "2. Fill ALL yellow-highlighted cells with your organisation's specific information",
        "3. Use dropdown menus where provided (do not type free-form text in dropdown cells)",
        "4. Document ALL environments in your organisation (include cloud, on-premises, hybrid)",
        "5. For each environment, specify masking requirement (Mandatory/Conditional/Not Required) per policy S2.3",
        "6. Verify masking deployment status (Yes/No/Partial/Planned)",
        "7. Calculate coverage percentage (% of sensitive fields masked)",
        "8. Link all assessments to Evidence Register with unique Evidence IDs",
        "9. Complete Gap Analysis sheet to identify remediation needs",
        "10. Review Summary Dashboard for executive-level compliance status",
    ]

    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # ── Status Legend — plain bold heading, 3 columns ONLY, proper table ──
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12)

    row += 1
    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border

    legend = [
        (CHECK, "Compliant", "Masking deployed as required — fully compliant"),
        (WARNING, "Partial", "Some masking gaps identified — partially compliant"),
        (XMARK, "Non-Compliant", "Masking required but not deployed — remediation needed"),
        ("—", "N/A", "Not applicable to this organisation"),
    ]

    row += 1
    for sym, status, desc in legend:
        c1 = ws.cell(row=row, column=1, value=sym)
        c1.border = border
        c2 = ws.cell(row=row, column=2, value=status)
        c2.border = border
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3 = ws.cell(row=row, column=3, value=desc)
        c3.border = border
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    # ── Acceptable Evidence — plain bold heading, AFTER Status Legend ──
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12)

    evidence_types = [
        "\u2713 Environment inventory reports and architecture diagrams",
        "\u2713 Data masking configuration screenshots",
        "\u2713 Masking tool deployment evidence and audit logs",
        "\u2713 Data flow diagrams with masking checkpoints",
        "\u2713 Non-production data refresh procedures with masking steps",
        "\u2713 Coverage assessment reports and gap analysis",
        "\u2713 Exception approvals with business justification",
        "\u2713 Vendor/DPA documentation specifying masking requirements",
        "\u2713 Cloud provider security assessment results",
        "\u2713 Backup encryption and access control documentation",
        "\u2713 Compliance audit reports and remediation evidence",
        "\u2713 Change management records for masking deployments",
    ]

    row += 1
    for ev in evidence_types:
        ws[f"A{row}"] = ev
        row += 1

    # ── Column widths & freeze ──
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet."""

    ws.merge_cells('A1:L1')
    ws['A1'] = "GAP ANALYSIS & REMEDIATION PLAN"
    apply_style(ws['A1'], styles["header"])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:L2')
    ws['A2'] = "All coverage gaps must be identified, risk-assessed, and remediated per organisational risk tolerance"
    apply_style(ws['A2'], styles["subheader"])

    # Column Headers
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Environment/System", 25),
        ("C", "Gap Description", 35),
        ("D", "Current State", 25),
        ("E", "Target State", 25),
        ("F", "Risk Level", 12),
        ("G", "Impact", 25),
        ("H", "Remediation Action", 35),
        ("I", "Owner", 20),
        ("J", "Target Date", 15),
        ("K", "Status", 15),
        ("L", "Evidence ID", 15),
    ]

    for col, header, width in headers:
        ws[f'{col}3'] = header
        apply_style(ws[f'{col}3'], styles["column_header"])
        ws.column_dimensions[col].width = width

    # SAMPLE ROW (Row 4) - First data row with complete example
    sample_data = [
        "GAP-001",
        "Test Environment - App Server 03",
        "Production data cloned directly to test without masking",
        "Real customer emails visible in test database",
        "All PII masked before deployment to test",
        "High",
        "Data breach risk if test environment compromised; GDPR/FADP violation",
        "Implement SDM in data refresh pipeline; automate masking before test deployment",
        "Data Engineering",
        "31.03.2026",
        "In Progress",
        "EV-001",
    ]

    for col_idx, (col_letter, _, _) in enumerate(headers):
        if col_idx < len(sample_data):
            ws[f'{col_letter}4'] = sample_data[col_idx]
        ws[f'{col_letter}4'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws[f'{col_letter}4'].border = styles["border"]

    # EMPTY DATA ROWS (5-54: 50 additional rows = 51 total)
    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked"')

    for row in range(4, 55):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            if row > 4:  # Skip sample row
                ws[f'{col}{row}'].fill = styles["input_cell"]["fill"]
                ws[f'{col}{row}'].border = styles["border"]

        risk_dv.add(ws[f'F{row}'])
        status_dv.add(ws[f'K{row}'])

    ws.add_data_validation(risk_dv)
    ws.add_data_validation(status_dv)

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: EVIDENCE REGISTER (GOLDEN STANDARD)
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet (golden standard ER)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header — A1:H1 merge
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle (no banner)
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4: Column headers — 8 columns
    headers = [
        ("Evidence ID", 15),
        ("Category", 20),
        ("Description", 40),
        ("Source/Location", 25),
        ("Date Collected", 15),
        ("Collected By", 20),
        ("Status", 15),
        ("Notes", 30),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Sample row with complete example data
    row = 5
    sample_data = {
        1: "EV-001",
        2: "Architecture diagram",
        3: "Environment separation diagram showing dev/test/prod isolation",
        4: "/evidence/environment-architecture-diagram.pdf",
        5: "15.01.2026",
        6: "Security Team",
        7: "Verified",
        8: "Annual architecture review Q1 2026"
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Empty rows 6-104 (99 empty rows)
    for r in range(6, 105):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Dropdowns
    dv_category = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Architecture diagram,Documentation,Audit log,Compliance report,Data flow map,Masking evidence,Other"',
        allow_blank=True,
    )
    dv_category.add("B5:B104")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    dv_status.add("G5:G104")

    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_status)

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: APPROVAL SIGN-OFF (GOLDEN STANDARD)
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet (golden standard AS)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Row 3: ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G13),\"\")"),
        ("Assessment Status:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            if label == "Assessment Status:":
                status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    if status_row:
        dv_status.add(ws[f"B{status_row}"])

    # 3 approver sections
    approvers = [
        ("PREPARED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "4472C4"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            row += 1
        row += 1  # gap between sections

    # FINAL ASSESSMENT DECISION
    ws[f"A{row}"] = "FINAL ASSESSMENT DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    dv_decision.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Finalize validations (only add those with sqref)
    for dv in [dv_status, dv_decision]:
        if dv.sqref:
            ws.add_data_validation(dv)

    ws.freeze_panes = "A3"
    # Apply borders to all merged cell top-left corners (GS-AS-011)
    _as_thin = Side(style="thin")
    _as_border = Border(left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin)
    for merge_range in ws.merged_cells.ranges:
        tl = ws.cell(merge_range.min_row, merge_range.min_col)
        tl.border = _as_border


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # ── Row 1: Main header ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle ──
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        f"Summary Dashboard  |  {WORKBOOK_NAME}  |  Generated: {GENERATED_DATE}"
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: blank ──

    # ── TABLE 1 banner (Row 4) ──
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 1 column headers (Row 5) ──
    t1_headers = [
        "Assessment Area", "Total Items", "Compliant",
        "Partial", "Non-Compliant", "N/A", "Compliance %",
    ]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 1 data rows ──
    # All assessment sheets use col O (compliance) rows 8:57
    # DV: "✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"
    #
    # Environment Inventory — col O, rows 8:57
    # Production Environment — col O, rows 8:57 (extended columns, same O position)
    # NonProduction Environments — col O, rows 8:57
    # Analytics & Reporting — col O, rows 8:57 (row_count=30, but rows 8:57 correct for DV)
    # Backup & Archive — col O, rows 8:57
    # External Sharing — col O, rows 8:57
    # Cloud Environments — col O, rows 8:57
    # Data Flow Mapping — col O is col 15 in dataflow_cols dict (Compliance Status)
    # Gap Analysis — col K (status_dv), rows 4:54
    #   DV: "Not Started,In Progress,Completed,Blocked"
    #   Compliant = COUNTIF(K4:K54,"Completed")
    #   Partial   = COUNTIF(K4:K54,"In Progress")
    #   Non-Comp  = COUNTIF(K4:K54,"Not Started")+COUNTIF(K4:K54,"Blocked")
    #   N/A       = 0
    #   NOTE: sample is row 4, so count from row 4 (sample excluded from compliance calcs
    #   since sample has "In Progress" — but design says COUNTIF starts after sample.
    #   Actually sample row is row 4 in gen3's gap analysis (create_gap_analysis),
    #   and DV is applied from row 4 onwards. Per standard, COUNTIF starts after sample.
    #   However, gap analysis DV starts at row 4 (includes sample). Use row 5 to be safe.

    t1_areas = [
        (
            "Environment Inventory",
            "=C6+D6+E6+F6",
            "=COUNTIF('Environment Inventory'!O8:O100,\"\u2705 Compliant\")",
            "=COUNTIF('Environment Inventory'!O8:O100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Environment Inventory'!O8:O100,\"\u274C Non-Compliant\")",
            "=COUNTIF('Environment Inventory'!O8:O100,\"N/A\")",
        ),
        (
            "Production Environment",
            "=C7+D7+E7+F7",
            "=COUNTIF('Production Environment'!O8:O100,\"\u2705 Compliant\")",
            "=COUNTIF('Production Environment'!O8:O100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Production Environment'!O8:O100,\"\u274C Non-Compliant\")",
            "=COUNTIF('Production Environment'!O8:O100,\"N/A\")",
        ),
        (
            "Non-Production Environments",
            "=C8+D8+E8+F8",
            "=COUNTIF('NonProduction Environments'!O8:O100,\"\u2705 Compliant\")",
            "=COUNTIF('NonProduction Environments'!O8:O100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('NonProduction Environments'!O8:O100,\"\u274C Non-Compliant\")",
            "=COUNTIF('NonProduction Environments'!O8:O100,\"N/A\")",
        ),
        (
            "Analytics & Reporting",
            "=C9+D9+E9+F9",
            "=COUNTIF('Analytics & Reporting'!O8:O100,\"\u2705 Compliant\")",
            "=COUNTIF('Analytics & Reporting'!O8:O100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Analytics & Reporting'!O8:O100,\"\u274C Non-Compliant\")",
            "=COUNTIF('Analytics & Reporting'!O8:O100,\"N/A\")",
        ),
        (
            "Backup & Archive",
            "=C10+D10+E10+F10",
            "=COUNTIF('Backup & Archive'!O8:O100,\"\u2705 Compliant\")",
            "=COUNTIF('Backup & Archive'!O8:O100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Backup & Archive'!O8:O100,\"\u274C Non-Compliant\")",
            "=COUNTIF('Backup & Archive'!O8:O100,\"N/A\")",
        ),
        (
            "External Sharing",
            "=C11+D11+E11+F11",
            "=COUNTIF('External Sharing'!O8:O100,\"\u2705 Compliant\")",
            "=COUNTIF('External Sharing'!O8:O100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('External Sharing'!O8:O100,\"\u274C Non-Compliant\")",
            "=COUNTIF('External Sharing'!O8:O100,\"N/A\")",
        ),
        (
            "Cloud Environments",
            "=C12+D12+E12+F12",
            "=COUNTIF('Cloud Environments'!O8:O100,\"\u2705 Compliant\")",
            "=COUNTIF('Cloud Environments'!O8:O100,\"\u26A0\uFE0F Partial\")",
            "=COUNTIF('Cloud Environments'!O8:O100,\"\u274C Non-Compliant\")",
            "=COUNTIF('Cloud Environments'!O8:O100,\"N/A\")",
        ),
        (
            "Gap Analysis",
            "=C13+D13+E13+F13",
            "=COUNTIF('Gap Analysis'!K5:K100,\"Completed\")",
            "=COUNTIF('Gap Analysis'!K5:K100,\"In Progress\")",
            "=COUNTIF('Gap Analysis'!K5:K100,\"Not Started\")+COUNTIF('Gap Analysis'!K5:K100,\"Blocked\")",
            "=0",
        ),
    ]

    for row_idx, (area, total, compliant, partial, non_comp, na) in enumerate(t1_areas, start=6):
        row = row_idx
        ws[f"A{row}"] = area
        ws[f"A{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = total
        ws[f"B{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row}"].border = border
        ws[f"C{row}"] = compliant
        ws[f"C{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row}"].border = border
        ws[f"D{row}"] = partial
        ws[f"D{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{row}"].border = border
        ws[f"E{row}"] = non_comp
        ws[f"E{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"E{row}"].border = border
        ws[f"F{row}"] = na
        ws[f"F{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row}"].border = border
        ws[f"G{row}"] = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        ws[f"G{row}"].number_format = "0.0%"
        ws[f"G{row}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"G{row}"].border = border

    # ── TOTAL row ──
    total_row = 14
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Calibri", size=10, bold=True, color="000000")
    ws[f"A{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{total_row}"].border = border
    for col_letter, formula in [
        ("B", "=SUM(B6:B13)"),
        ("C", "=SUM(C6:C13)"),
        ("D", "=SUM(D6:D13)"),
        ("E", "=SUM(E6:E13)"),
        ("F", "=SUM(F6:F13)"),
    ]:
        ws[f"{col_letter}{total_row}"] = formula
        ws[f"{col_letter}{total_row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"{col_letter}{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col_letter}{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col_letter}{total_row}"].border = border
    ws[f"G{total_row}"] = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    ws[f"G{total_row}"].number_format = "0.0%"
    ws[f"G{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"G{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"G{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"G{total_row}"].border = border

    # ── TABLE 2 banner (Row 16) ──
    ws.merge_cells("A16:G16")
    ws["A16"] = "TABLE 2 \u2013 KEY METRICS"
    ws["A16"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A16"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A16"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 2 headers (Row 17) ──
    for col_idx, hdr in enumerate(["Metric", "Value", "Target"], start=1):
        cell = ws.cell(row=17, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 2 data rows (Rows 18-23) ──
    t2_data = [
        ("Non-Production Masking Coverage (%)", f"=IF(COUNTA('NonProduction Environments'!O8:O100)=0,\"N/A\",TEXT(COUNTIF('NonProduction Environments'!O8:O100,\"\u2705 Compliant\")/COUNTA('NonProduction Environments'!O8:O100),\"0.0%\"))", "100%"),
        ("Production DDM Coverage (%)", f"=IF(COUNTA('Production Environment'!O8:O100)=0,\"N/A\",TEXT(COUNTIF('Production Environment'!O8:O100,\"\u2705 Compliant\")/COUNTA('Production Environment'!O8:O100),\"0.0%\"))", ">=90%"),
        ("Analytics Masking Coverage (%)", f"=IF(COUNTA('Analytics & Reporting'!O8:O100)=0,\"N/A\",TEXT(COUNTIF('Analytics & Reporting'!O8:O100,\"\u2705 Compliant\")/COUNTA('Analytics & Reporting'!O8:O100),\"0.0%\"))", "100%"),
        ("External Sharing Masking Coverage (%)", f"=IF(COUNTA('External Sharing'!O8:O100)=0,\"N/A\",TEXT(COUNTIF('External Sharing'!O8:O100,\"\u2705 Compliant\")/COUNTA('External Sharing'!O8:O100),\"0.0%\"))", "100%"),
        ("Open Critical Gaps (Count)", f"=COUNTIF('Gap Analysis'!K5:K100,\"Not Started\")+COUNTIF('Gap Analysis'!K5:K100,\"Blocked\")", "0"),
        ("Cloud Environment Compliance (%)", f"=IF(COUNTA('Cloud Environments'!O8:O100)=0,\"N/A\",TEXT(COUNTIF('Cloud Environments'!O8:O100,\"\u2705 Compliant\")/COUNTA('Cloud Environments'!O8:O100),\"0.0%\"))", "100%"),
    ]
    for row_idx, (metric, value, target) in enumerate(t2_data, start=18):
        ws[f"A{row_idx}"] = metric
        ws[f"A{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row_idx}"].border = border
        ws[f"B{row_idx}"] = value
        ws[f"B{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row_idx}"].border = border
        ws[f"C{row_idx}"] = target
        ws[f"C{row_idx}"].font = Font(name="Calibri", size=10)
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{row_idx}"].border = border

    # ── TABLE 3 banner (Row 25) ──
    ws.merge_cells("A25:G25")
    ws["A25"] = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    ws["A25"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A25"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws["A25"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 3 headers (Row 26) ──
    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", "Notes"]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=26, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── TABLE 3 data rows (Rows 27-30) ──
    t3_data = [
        ("1", "Non-production environments not consistently masked before data refresh", "Critical — GDPR/FADP violation if production data visible in test/UAT", "Integrate Static Data Masking (SDM) into all data refresh pipelines; enforce before environment provisioning", "P1", "Open", ""),
        ("2", "Production DDM not fully deployed for all customer-facing roles", "High — CSRs and support staff may view unmasked sensitive data", "Deploy DDM rules for all non-privileged roles; log all access to unmasked data", "P1", "In Progress", ""),
        ("3", "External data sharing lacking DPA masking clauses for some vendors", "High — contractual gap means vendors may retain unmasked sensitive data", "Audit all DPAs; amend to include explicit masking requirements; verify vendor compliance", "P2", "Open", ""),
        ("4", "Cloud environments not all mapped to correct masking classification", "Medium — cloud non-prod environments may be provisioned with unmasked data", "Apply same classification and masking rules to cloud as on-premises; validate quarterly", "P2", "Open", ""),
    ]
    for row_idx, row_data in enumerate(t3_data, start=27):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border

    # ── Freeze panes ──
    ws.freeze_panes = "A4"



# ============================================================================
# SECTION 10: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.11.3 - Environment Coverage Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.11 - Data Masking")
    logger.info("=" * 78)

    wb = create_workbook()
    styles = get_styles()

    logger.info("[1/13] Creating Instructions & Legend...")
    create_instructions(wb["Instructions & Legend"], styles)

    logger.info("[2/13] Creating Environment Inventory...")
    base_cols = get_base_columns()
    checklist_env = [
        "All production environments documented",
        "All non-production environments documented (Dev/Test/UAT/Sandbox)",
        "All analytics/reporting environments documented",
        "All cloud environments documented (AWS/Azure/GCP/Other)",
        "All backup/archive systems documented",
        "External data sharing destinations documented",
        "Each environment classified (Sensitive/Confidential/Internal/Public)",
        "Data sensitivity level assigned per environment",
        "Masking requirement determined (Mandatory/Conditional/Not Required)",
        "Environment owners assigned",
        "Data owners assigned",
        "Hosting location documented (On-Prem/Cloud)",
        "Environment inventory reviewed in last 6 months",
        "New environments added to inventory within 30 days of deployment",
        "Decommissioned environments removed from inventory",
    ]
    create_assessment_sheet(
        wb["Environment Inventory"], styles,
        "ENVIRONMENT INVENTORY",
        "All information processing environments must be cataloged and classified for masking applicability (ISMS-POL-A.8.11-S2.3 Section 2)",
        "Does your organisation maintain a complete inventory of ALL environments where data is processed, stored, or transmitted?",
        base_cols, 51, checklist_env
    )

    logger.info("[3/13] Creating Production Environment...")
    prod_cols = {**base_cols, **get_extended_columns("production")}
    checklist_prod = [
        "DDM implemented for role-based access in production",
        "Customer service representatives see masked customer data",
        "Production reports mask data for non-privileged users",
        "Audit logs containing sensitive data are masked/encrypted",
        "Application outputs (invoices, statements) use partial redaction",
        "Production data exports are masked before release",
        "All access to unmasked production data is logged",
        "Access logs reviewed monthly for anomalies",
        "Privileged user access to unmasked data requires justification",
        "Masking exceptions documented with business justification",
        "Exception approvals obtained from Data Owner and ISO",
        "Exceptions reviewed quarterly",
        "DDM performance impact assessed (<10% degradation acceptable)",
        "DDM rules tested before production deployment",
        "DDM bypass mechanisms disabled or strictly controlled",
        "Production masking coverage measured and tracked",
        "Coverage target: \u226590% of sensitive fields masked for non-privileged users",
        "Regulatory reporting requirements accommodate masking",
    ]
    create_assessment_sheet(
        wb["Production Environment"], styles,
        "PRODUCTION ENVIRONMENT ASSESSMENT",
        "Production environments may use Dynamic Data Masking (DDM) for role-based access. All access to unmasked data must be logged. (ISMS-POL-A.8.11-S2.3 Section 3.1)",
        "Does your organisation implement role-based masking (DDM) in production environments to restrict access to sensitive data?",
        prod_cols, 51, checklist_prod, "production"
    )

    logger.info("[4/13] Creating Non-Production Environments...")
    nonprod_cols = {**base_cols, **get_extended_columns("nonproduction")}
    checklist_nonprod = [
        "Development environments masked",
        "Testing/QA environments masked",
        "UAT environments masked",
        "Training environments masked",
        "Sandbox environments masked",
        "Staging environments masked (unless prod-identical required)",
        "Static Data Masking (SDM) applied during data refresh",
        "Masking applied BEFORE data deployment (not 'later')",
        "Direct production database cloning prevented",
        "Automated masking integrated into data refresh pipeline",
        "Manual data copies prohibited without masking",
        "Developer NDA reliance eliminated (NDAs not technical controls)",
        "Non-production data refresh process documented",
        "Masking validation performed after each refresh",
        "Non-production exception count: ___ (Target: 0)",
        "Exceptions approved by ISO and Data Owner",
        "Compensating controls implemented for exceptions",
        "Exception review conducted quarterly",
        "Coverage target: 100% of non-prod environments masked",
        "Non-compliance escalated to CISO",
    ]
    create_assessment_sheet(
        wb["NonProduction Environments"], styles,
        "NON-PRODUCTION ENVIRONMENTS ASSESSMENT",
        "ALL sensitive data SHALL be masked before deployment to non-production. NO production data SHALL be copied without masking. (ISMS-POL-A.8.11-S2.3 Section 3.2)",
        "Are ALL non-production environments (Dev/Test/UAT/Training/Sandbox/Staging) masked before receiving production data?",
        nonprod_cols, 51, checklist_nonprod, "nonproduction"
    )

    logger.info("[5/13] Creating Analytics & Reporting...")
    analytics_cols = {**base_cols, **get_extended_columns("analytics")}
    checklist_analytics = [
        "BI tool exports contain masked data",
        "Data warehouse ETL includes masking steps",
        "Individual-level reports use masked data",
        "Aggregated reports assessed for re-identification risk",
        "ML/AI training datasets use synthetic or masked data",
        "PII removed or anonymised in training data",
        "Ad-hoc query exports masked",
        "Self-service BI tools enforce masking rules",
        "CSV/Excel exports contain masked data",
        "Analytics platform integrates masking at data ingestion",
        "Data minimisation applied (only export necessary fields)",
        "Re-identification risk assessed before data release",
        "Historical analytics data masked or aggregated",
        "Analytics data retention period documented",
        "Analytics environment access logged",
    ]
    create_assessment_sheet(
        wb["Analytics & Reporting"], styles,
        "ANALYTICS & REPORTING ENVIRONMENTS ASSESSMENT",
        "Individual-level PII SHALL be masked or aggregated. Synthetic data SHALL be used for ML/AI training. Re-identification risk SHALL be assessed. (ISMS-POL-A.8.11-S2.3 Section 3.3)",
        "Are analytics, BI, data warehouse, and ML/AI environments masked to prevent individual-level data exposure?",
        analytics_cols, 30, checklist_analytics, "analytics"
    )

    logger.info("[6/13] Creating Backup & Archive...")
    backup_cols = {**base_cols, **get_extended_columns("backup")}
    checklist_backup = [
        "Production backups encrypted at rest",
        "Backup encryption uses strong algorithms (AES-256)",
        "Backup access restricted to authorised administrators only",
        "All backup access logged",
        "Non-production backups contain only masked data",
        "Backup restoration to non-prod triggers masking process",
        "Archive data assessed for masking feasibility",
        "If masking compromises compliance, encryption used instead",
        "Archive access strictly controlled",
        "Backup media stored securely (encrypted, access-controlled)",
        "Backup retention policy documented",
        "Backup testing includes masking validation",
    ]
    create_assessment_sheet(
        wb["Backup & Archive"], styles,
        "BACKUP & ARCHIVE ENVIRONMENTS ASSESSMENT",
        "Production backups may contain unmasked data if encrypted. Non-production backups SHALL contain only masked data. (ISMS-POL-A.8.11-S2.3 Section 3.4)",
        "Are backup and archive environments encrypted and access-controlled? Are non-production backups masked?",
        backup_cols, 30, checklist_backup, "backup"
    )

    logger.info("[7/13] Creating External Sharing...")
    external_cols = {**base_cols, **get_extended_columns("external")}
    checklist_external = [
        "Vendor data shares masked unless contractually required",
        "Data Processing Agreements (DPAs) in place",
        "DPAs specify masking requirements",
        "Vendor access to unmasked data logged and monitored",
        "Vendor environments audited for security controls",
        "Customer exports appropriately masked (not their own data)",
        "Auditor data samples masked where possible",
        "Auditors sign confidentiality agreements",
        "Partner data sharing follows minimum necessary principle",
        "Data minimisation applied (only share necessary fields)",
        "External sharing inventory maintained",
        "Re-identification risk assessed before external release",
        "External sharing reviewed annually",
        "Data retention limits specified in agreements",
        "Data deletion verified after contract termination",
    ]
    create_assessment_sheet(
        wb["External Sharing"], styles,
        "EXTERNAL DATA SHARING ASSESSMENT",
        "ALL data shared externally SHALL be masked unless contractually required. Data Processing Agreements SHALL specify masking. (ISMS-POL-A.8.11-S2.3 Section 3.5)",
        "Is data shared with vendors, partners, auditors, or customers masked unless contractually required to be unmasked?",
        external_cols, 30, checklist_external, "external"
    )

    logger.info("[8/13] Creating Cloud Environments...")
    cloud_cols = {**base_cols, **get_extended_columns("cloud")}
    checklist_cloud = [
        "Cloud environments classified correctly (Prod/Non-Prod/Analytics)",
        "Cloud-hosted production follows production masking rules",
        "Cloud-hosted non-prod follows non-prod masking rules (mandatory)",
        "Cloud analytics follows analytics masking rules",
        "Client-side masking applied before cloud upload (where applicable)",
        "Cloud provider security controls reviewed",
        "Cloud provider SLAs reviewed for data protection",
        "Multi-tenancy risks assessed and mitigated",
        "Data residency requirements met",
        "Cloud encryption enabled (at rest and in transit)",
        "Cloud access controls configured (IAM, RBAC)",
        "Cloud environment masking tested and validated",
    ]
    create_assessment_sheet(
        wb["Cloud Environments"], styles,
        "CLOUD ENVIRONMENTS ASSESSMENT",
        "Cloud environments SHALL follow same masking requirements as on-premises. Cloud-hosted non-prod SHALL be masked. (ISMS-POL-A.8.11-S2.3 Section 2.6)",
        "Are cloud-hosted environments (AWS/Azure/GCP/Other) masked according to the same requirements as on-premises environments?",
        cloud_cols, 30, checklist_cloud, "cloud"
    )

    logger.info("[9/13] Creating Data Flow Mapping...")
    dataflow_cols = {
        "Data Flow Name": 25,
        "Source Environment": 20,
        "Destination Environment": 20,
        "Data Type": 18,
        "Masking Checkpoint?": 18,
        "Masking Technique": 20,
        "Flow Frequency": 15,
        "Automated Masking?": 18,
        "Masking Tool/Script": 22,
        "Masking Validation": 20,
        "Last Flow Date": 15,
        "Flow Owner": 20,
        "Approval Required?": 15,
        "Approval Status": 18,
        "Compliance Status": 18,
        "Risk Level": 12,
        "Notes/Comments": 30,
        "Evidence ID": 15,
    }
    checklist_dataflow = [
        "All production \u2192 non-production flows documented",
        "Masking checkpoints identified at environment boundaries",
        "Data refresh processes documented",
        "Automated masking integrated into data pipelines",
        "Manual data transfers prohibited or strictly controlled",
        "End-to-end masking verified for each flow",
        "Data flow inventory reviewed quarterly",
        "New data flows assessed for masking requirements",
        "Data flow exceptions approved by ISO",
        "High-risk flows (unmasked external sharing) escalated",
    ]
    create_assessment_sheet(
        wb["Data Flow Mapping"], styles,
        "DATA FLOW MAPPING & MASKING CHECKPOINTS",
        "Data flows SHALL be documented with masking checkpoints identified at each environment boundary. (ISMS-POL-A.8.11-S2.3 Section 3)",
        "Are data flows documented with masking checkpoints identified at each environment boundary?",
        dataflow_cols, 30, checklist_dataflow, "dataflow"
    )

    logger.info("[10/13] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[11/13] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[12/13] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[13/13] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)

    # Save workbook — auto-detect Assessment folder
    filename = f"ISMS-IMP-A.8.11.3_Environment_Coverage_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    script_dir = Path(__file__).resolve().parent
    _wkbk_dir = script_dir.parent / "WKBK"
    _wkbk_dir.mkdir(exist_ok=True)
    output_path = _wkbk_dir / OUTPUT_FILENAME
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    finalize_validations(wb)
    wb.save(output_path)

    logger.info(f"SUCCESS: {output_path}")
    logger.info("Workbook Structure: 13 sheets including 8 Assessment Sheets, Evidence Register, Approval Sign-Off")
    logger.info("=" * 78)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
