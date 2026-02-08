#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-21-22.6 - Network Security Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Dashboard: Executive Compliance Overview (Consolidates WB1-5)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dashboard requirements, reporting standards, and
management expectations.

Key customization areas:
1. Compliance scoring thresholds (match your risk appetite)
2. Dashboard metrics and KPIs (based on management requirements)
3. Gap prioritization criteria (aligned with your risk assessment)
4. Trend analysis periods (adapted to your review cycles)
5. Executive summary content (tailored to your stakeholders)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-21-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates the master compliance dashboard that consolidates findings
from all five network security assessment workbooks (WB1-5) into an executive-level
view for management oversight and decision-making.

**Purpose:**
Consolidates findings from all five network security assessment workbooks into
executive-level compliance dashboard, providing management visibility into overall
network security posture, compliance status, and remediation priorities across
ISO 27001:2022 Controls A.8.20, A.8.21, and A.8.22.

**Assessment Scope:**
- Aggregate compliance scoring (A.8.20, A.8.21, A.8.22 separate + combined)
- Executive summary for management
- Device inventory and security status summary
- Network services security posture
- Segmentation effectiveness metrics
- Controls coverage analysis
- Gap consolidation and prioritization
- Remediation roadmap and timelines
- Trend analysis across assessments
- Evidence summary for audit readiness

**Generated Workbook Structure:**
1. Instructions - Dashboard usage guidance and data source information
2. Executive_Summary - High-level findings for C-level and management
3. Overall_Compliance - Aggregate compliance across A.8.20, A.8.21, A.8.22
4. A820_Network_Security - Control A.8.20 compliance dashboard
5. A821_Network_Services - Control A.8.21 compliance dashboard
6. A822_Network_Segregation - Control A.8.22 compliance dashboard
7. Device_Security_Summary - Device inventory and hardening status
8. Services_Security_Summary - Network services security posture
9. Segmentation_Summary - Network segmentation effectiveness
10. Controls_Coverage_Summary - Unified controls coverage analysis
11. Gap_Consolidation - All gaps from WB1-5 prioritized
12. Remediation_Roadmap - Prioritized action items with timelines
13. Compliance_Trends - Historical compliance trends (if data available)
14. Evidence_Summary - Audit evidence status across all assessments
15. Approval_Sign_Off - Executive review and approval workflow

**Key Features:**
- Executive-focused data visualization
- Compliance scoring with RAG (Red/Amber/Green) status
- Automated gap consolidation from all workbooks
- Protected formulas with external workbook references
- Trend analysis and historical tracking
- Risk-based gap prioritization
- Evidence completeness tracking
- Multi-level approval workflow
- Integration with WB1-5 via external formulas

**Integration:**
This dashboard is the final deliverable of the network security assessment
framework, consolidating data from WB1 (Infrastructure Inventory), WB2 (Device
Security), WB3 (Network Services), WB4 (Segmentation), and WB5 (Controls
Coverage) into a single executive view.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
PREREQUISITES
--------------------------------------------------------------------------------

**CRITICAL: This dashboard generator requires:**
1. All 5 assessment workbooks generated (Scripts 1-5)
2. Workbooks normalized using normalize_a820_assessments.py
3. Normalized files named: ISMS-IMP-A.8.20-21-22.X.xlsx (where X = 1-5)
4. All normalized files in same directory as this script

Expected Normalized Files (in same directory):
    ISMS-IMP-A.8.20-21-22.S1.xlsx  (Infrastructure Inventory)
    ISMS-IMP-A.8.20-21-22.S2.xlsx  (Device Security Assessment)
    ISMS-IMP-A.8.20-21-22.S3.xlsx  (Network Services Catalog)
    ISMS-IMP-A.8.20-21-22.S4.xlsx  (Segmentation Matrix)
    ISMS-IMP-A.8.20-21-22.S5.xlsx  (Controls Coverage Matrix)

**Dashboard Consolidation:**
The dashboard uses external formulas to read data from normalized assessment
workbooks. Do NOT modify the normalized filenames or move files after dashboard
generation, as this will break formula references.

**Data Refresh:**
When assessment workbooks are updated:
1. Re-normalize updated workbooks (Script 7)
2. Open dashboard and refresh all data connections
3. Verify compliance scores updated correctly
4. Update "Last Updated" date in dashboard

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_6_compliance_dashboard.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_6_compliance_dashboard.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_6_compliance_dashboard.py --date 20250124
    
    # Generate with custom organization name
    python3 generate_a820_6_compliance_dashboard.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organization name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Verify all prerequisite workbooks (WB1-5) are normalized
    2. Open dashboard and verify external formula references work
    3. Review executive summary for accuracy
    4. Validate compliance scores against source workbooks
    5. Review gap consolidation and prioritization
    6. Update remediation roadmap with realistic timelines
    7. Review evidence summary for completeness
    8. Obtain executive approvals
    9. Distribute dashboard to stakeholders
    10. Schedule periodic dashboard updates (quarterly recommended)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    Dashboard (Executive Compliance Overview - Consolidates WB1-5)
Primary Control:      A.8.20-22 (Unified Network Security Framework)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-21-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-21-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-21-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-21-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-21-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-21-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-21-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-21-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-21-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-21-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-21-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-21-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalization)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full dashboard framework per ISMS-IMP-A.8.20-21-22 specification
    - Consolidates data from all five assessment workbooks
    - Provides executive-level compliance reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Dashboard Purpose:**
This dashboard is designed for executive and management consumption. Focus on:
- High-level compliance status (RAG ratings)
- Critical gaps requiring immediate attention
- Clear remediation priorities and timelines
- Business risk context (not just technical details)

Avoid overwhelming executives with technical minutiae.

**Compliance Scoring:**
Customize scoring thresholds to match your organization's risk appetite:
- GREEN (Compliant): 90-100% compliance
- AMBER (Partial): 70-89% compliance
- RED (Non-Compliant): <70% compliance

Adjust thresholds based on your risk tolerance and regulatory requirements.

**Technology Diversity:**
This dashboard consolidates data across all network architectures:
- Traditional networks
- Software-Defined Networks
- Cloud environments
- Hybrid architectures

Ensure dashboard metrics are meaningful across all environments.

**Audit Considerations:**
This dashboard serves as audit evidence per ISO 27001:2022 requirements.
Auditors will use this to:
- Assess overall network security posture
- Verify management oversight and accountability
- Review gap remediation progress
- Validate evidence completeness

Ensure dashboard data is accurate and traceable to source assessments.

**Data Protection:**
Dashboard contains high-level security posture information including:
- Overall compliance status and gaps
- Critical vulnerabilities and remediation plans
- Network security architecture summary

Handle in accordance with your organization's data classification policies.
Distribute to authorized executives and senior management only.

**Maintenance:**
Update dashboard:
- Monthly: After significant network changes or gap remediation
- Quarterly: Routine dashboard refresh (recommended)
- Annually: Complete reassessment with full dashboard regeneration
- Ad-hoc: After security incidents or audit requests

**Quality Assurance:**
Validate dashboard accuracy by:
- Cross-checking compliance scores with source workbooks
- Verifying gap counts match consolidated findings
- Reviewing remediation roadmap for realism
- Peer review by CISO or security management
- Stakeholder feedback on dashboard usability

**Regulatory Alignment:**
Ensure dashboard reporting meets applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 reporting requirements
- Healthcare: HIPAA security posture reporting
- Finance: Regional banking compliance reporting
- Government: Jurisdiction-specific reporting mandates

Customize dashboard content to address regulatory reporting needs.

**Integration Points:**
This dashboard provides executive view across all network security controls:
- A.8.20 (Network Security): Device and infrastructure security
- A.8.21 (Network Services): Network services security
- A.8.22 (Segregation): Network segmentation effectiveness

Plus integration status with:
- A.8.15 (Logging): Logging implementation across network
- A.8.16 (Monitoring): Monitoring coverage across network

**Common Pitfalls to Avoid:**
1. **Broken References**: Moving normalized workbooks breaks external formulas
2. **Stale Data**: Not refreshing dashboard after assessment updates
3. **Too Technical**: Overwhelming executives with technical details
4. **No Context**: Presenting numbers without business risk context
5. **Unrealistic Remediation**: Setting unachievable timelines
6. **No Trends**: Missing historical trend data for progress tracking
7. **No Actions**: Identifying gaps without clear remediation plans
8. **No Accountability**: Not assigning ownership for remediation

**Dashboard Best Practices:**

**Executive Summary Content:**
- Overall network security posture (one paragraph)
- Key compliance metrics (A.8.20, A.8.21, A.8.22 scores)
- Critical gaps requiring immediate attention (top 5)
- Remediation progress since last assessment
- Resource requirements for gap closure

**Gap Prioritization Criteria:**
1. **Critical**: Gaps exposing organization to immediate significant risk
2. **High**: Gaps in important areas or partial control failures
3. **Medium**: Gaps in less critical areas or minor control weaknesses
4. **Low**: Enhancement opportunities or minimal risk

**Remediation Roadmap Structure:**
For each gap:
- Gap description and affected systems/zones
- Business risk and impact
- Remediation action required
- Estimated effort (person-hours or budget)
- Timeline (target completion date)
- Owner/responsible party
- Status (Not Started / In Progress / Complete)

**Compliance Trends:**
If historical data available:
- Compliance score trends over time (quarterly)
- Gap closure rate
- New gaps identified vs. gaps closed
- Time to close critical gaps
- Budget spent on network security improvements

**Stakeholder Communication:**
- CISO/Security Management: Monthly dashboard review
- Executive Management: Quarterly dashboard briefing
- Board of Directors: Annual network security posture report
- Auditors: On-demand access for audit evidence

**External Formula Management:**

The dashboard uses external formulas like:
```
='[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!$G$5
```

**Important:**
- Keep normalized workbooks in same directory as dashboard
- Don't rename normalized workbooks
- Don't move files between directories
- When distributing dashboard, include all normalized workbooks
- Document file dependencies for users

**Data Refresh Process:**
1. Open dashboard workbook
2. Excel will prompt about external links
3. Click "Update" to refresh all data
4. Verify all formulas calculate correctly
5. Save updated dashboard

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.20-21-22.6"
WORKBOOK_NAME = "Network Security Compliance Dashboard"
CONTROL_ID = "A.8.20-22"
CONTROL_NAME = "Networks Security"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

import sys
import os

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference
    from openpyxl.formatting.rule import CellIsRule
except ImportError as e:
    logger.error(f"❌ ERROR: Required library not found: {e}")
    logger.info("📦 Install required libraries: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================

WORKBOOK_NAME = "Network Security Compliance Dashboard"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Expected normalized workbook files (must be in same directory)
REQUIRED_WORKBOOKS = [
    "ISMS-IMP-A.8.20-21-22.S1.xlsx",
    "ISMS-IMP-A.8.20-21-22.S2.xlsx",
    "ISMS-IMP-A.8.20-21-22.S3.xlsx",
    "ISMS-IMP-A.8.20-21-22.S4.xlsx",
    "ISMS-IMP-A.8.20-21-22.S5.xlsx",
]

# Workbook metadata
WORKBOOK_INFO = {
    "ISMS-IMP-A.8.20-21-22.S1.xlsx": {
        "short": "WB1",
        "title": "Infrastructure Inventory",
        "control": "A.8.20",
        "description": "Network devices and infrastructure"
    },
    "ISMS-IMP-A.8.20-21-22.S2.xlsx": {
        "short": "WB2",
        "title": "Device Security",
        "control": "A.8.20",
        "description": "Device hardening compliance"
    },
    "ISMS-IMP-A.8.20-21-22.S3.xlsx": {
        "short": "WB3",
        "title": "Services Catalog",
        "control": "A.8.21",
        "description": "Network services security"
    },
    "ISMS-IMP-A.8.20-21-22.S4.xlsx": {
        "short": "WB4",
        "title": "Segmentation Matrix",
        "control": "A.8.22",
        "description": "Network segmentation"
    },
    "ISMS-IMP-A.8.20-21-22.S5.xlsx": {
        "short": "WB5",
        "title": "Controls Coverage",
        "control": "All",
        "description": "Security controls coverage"
    },
}


# ============================================================================
# SECTION 2: PREREQUISITE CHECKS
# ============================================================================

def check_prerequisites():
    """Verify all required normalized workbooks are present."""
    current_dir = Path.cwd()
    missing_files = []
    
    logger.info("\n🔍 Checking for required normalized workbooks...\n")
    
    for filename in REQUIRED_WORKBOOKS:
        filepath = current_dir / filename
        if filepath.exists():
            logger.info(f"  ✅ Found: {filename}")
        else:
            logger.info(f"  ❌ Missing: {filename}")
            missing_files.append(filename)
    
    if missing_files:
        logger.info("\n" + "=" * 80)
        logger.error("❌ ERROR: Missing Required Workbooks")
        logger.info("=" * 80)
        logger.info("\nThe following normalized workbooks are required but not found:\n")
        for filename in missing_files:
            logger.info(f"  • {filename}")
        logger.info("\nACTION REQUIRED:")
        logger.info("  1. Generate missing assessment workbooks (Scripts 1-5)")
        logger.info("  2. Run normalization script:")
        logger.info("     python3 normalize_network_security_assessments.py")
        logger.info("  3. Ensure normalized files are in current directory")
        logger.info("  4. Re-run this dashboard generator")
        logger.info("\n" + "=" * 80 + "\n")
        return False
    
    logger.info("\n✅ All required workbooks found\n")
    return True


# ============================================================================
# SECTION 3: STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "metric_label": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "metric_value": {
            "font": Font(name="Calibri", size=11, bold=False),
            "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=11, bold=True, color="006100"),
        },
        "noncompliant_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=11, bold=True, color="9C0006"),
        },
        "partial_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=11, bold=False),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=11, bold=False),
        },
        "info_box": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "exec_summary": {
            "font": Font(name="Calibri", size=12, bold=False),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 11),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions sheet explaining dashboard usage."""
    
    ws.title = "Instructions"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Network Security Compliance Dashboard - Instructions"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30
    
    # Dashboard Information
    ws.merge_cells("A3:B3")
    ws["A3"] = "Dashboard Information"
    apply_style(ws["A3"], styles["header"])
    
    info_data = [
        ("Dashboard Name:", WORKBOOK_NAME),
        ("Generated:", GENERATED_DATE),
        ("Version:", "1.0"),
        ("Controls:", "ISO 27001:2022 A.8.20, A.8.21, A.8.22"),
        ("Purpose:", "Consolidated compliance view across all network security assessments"),
        ("Data Sources:", "5 normalized assessment workbooks (WB1-WB5)"),
    ]
    
    row = 4
    for label, value in info_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Data Sources
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Data Sources (Linked Workbooks)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Workbook"
    ws["B" + str(row)] = "Title"
    ws["C" + str(row)] = "Control"
    ws["D" + str(row)] = "Description"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    for filename in REQUIRED_WORKBOOKS:
        info = WORKBOOK_INFO[filename]
        ws[f"A{row}"] = info["short"]
        ws[f"B{row}"] = info["title"]
        ws[f"C{row}"] = info["control"]
        ws[f"D{row}"] = info["description"]
        for col in ["A", "B", "C", "D"]:
            apply_style(ws[f"{col}{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Dashboard Sheets
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Dashboard Sheet Overview"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    sheet_info = [
        ("Executive_Summary", "High-level findings for management"),
        ("Overall_Compliance", "Aggregate compliance across A.8.20/21/22"),
        ("WB1_Infrastructure", "Device inventory summary (from WB1)"),
        ("WB2_Device_Security", "Device hardening compliance (from WB2)"),
        ("WB3_Services", "Network services security (from WB3)"),
        ("WB4_Segmentation", "Network segmentation effectiveness (from WB4)"),
        ("WB5_Controls_Coverage", "Controls coverage summary (from WB5)"),
        ("Gap_Consolidation", "All gaps from WB1-5 prioritized"),
        ("Action_Items", "Prioritized remediation roadmap"),
        ("Management_Dashboard", "Visual executive view with charts"),
    ]
    
    for sheet_name, description in sheet_info:
        ws[f"A{row}"] = sheet_name
        ws[f"B{row}"] = description
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Important Notes
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "⚠ IMPORTANT NOTES"
    apply_style(ws[f"A{row}"], styles["header"])
    
    notes = [
        "• This dashboard uses EXTERNAL LINKS to normalized workbooks (WB1-WB5)",
        "• All 5 workbooks must be in the SAME DIRECTORY as this dashboard",
        "• If you move files, links will break - use 'Edit Links' to update",
        "• Formulas automatically update when source workbooks are modified",
        "• For static snapshot, use 'Break Links' to convert to values",
        "• Red/yellow highlights indicate critical gaps requiring immediate attention",
    ]
    
    row += 1
    for note in notes:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = note
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 50
    for col in ["E", "F"]:
        ws.column_dimensions[col].width = 3


# ============================================================================
# SECTION 5: SHEET 2 - EXECUTIVE SUMMARY
# ============================================================================

def create_executive_summary_sheet(ws, styles):
    """Create Executive Summary for management."""
    
    ws.title = "Executive_Summary"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Network Security Compliance - Executive Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30
    
    # Assessment Overview
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Assessment Overview"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    overview_text = [
        f"Assessment Date: {GENERATED_DATE}",
        "",
        "This executive summary consolidates findings from comprehensive network security assessments",
        "covering ISO 27001:2022 controls A.8.20 (Networks Security), A.8.21 (Security of Network Services),",
        "and A.8.22 (Segregation of Networks).",
        "",
        "The assessment examined:",
        "  • Network infrastructure and device inventory (WB1)",
        "  • Device hardening compliance against CIS Benchmarks (WB2)",
        "  • Network services security (DNS, DHCP, NTP, Proxy, etc.) (WB3)",
        "  • Network segmentation architecture and effectiveness (WB4)",
        "  • Security controls coverage across all zones (WB5)",
    ]
    
    for line in overview_text:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = line
        apply_style(ws[f"A{row}"], styles["exec_summary"])
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Key Metrics Summary
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Key Metrics Summary"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    ws["C" + str(row)] = "Source"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    # Link to Overall_Compliance sheet metrics
    row += 1
    metrics = [
        ("Total Network Devices", "=Overall_Compliance!B5", "WB1"),
        ("Device Hardening Compliance", "=Overall_Compliance!B6", "WB2"),
        ("Network Services Cataloged", "=Overall_Compliance!B7", "WB3"),
        ("Security Zones Defined", "=Overall_Compliance!B8", "WB4"),
        ("Segmentation Test Pass Rate", "=Overall_Compliance!B9", "WB4"),
        ("Average Controls Coverage", "=Overall_Compliance!B10", "WB5"),
        ("Total Critical Gaps", "=Overall_Compliance!B11", "All WBs"),
    ]
    
    for metric, formula, source in metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        ws[f"C{row}"] = source
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        row += 1
    
    # Key Findings
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Key Findings"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    findings_text = [
        "STRENGTHS:",
        "  • [To be completed based on assessment results]",
        "  • [Example: Strong network segmentation with clearly defined security zones]",
        "  • [Example: Comprehensive device inventory maintained]",
        "",
        "AREAS FOR IMPROVEMENT:",
        "  • [To be completed based on gap analysis]",
        "  • [Example: Device hardening compliance below target (see WB2)]",
        "  • [Example: Several critical services lack redundancy (see WB3)]",
        "  • [Example: Missing defense-in-depth for critical zones (see WB5)]",
        "",
        "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION:",
        "  • [List from Gap_Consolidation sheet - Critical severity only]",
        "  • See Gap_Consolidation sheet for complete list",
    ]
    
    for line in findings_text:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = line
        apply_style(ws[f"A{row}"], styles["exec_summary"])
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Recommendations
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Priority Recommendations"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    recommendations_text = [
        "IMMEDIATE (Critical Priority):",
        "  1. Remediate all Critical gaps (see Gap_Consolidation sheet)",
        "  2. Implement defense-in-depth for critical zones (Datacenter, Management)",
        "  3. Address device hardening gaps for critical infrastructure",
        "",
        "SHORT-TERM (High Priority):",
        "  1. Improve network services security (DNS, DHCP, NTP hardening)",
        "  2. Enhance network segmentation effectiveness",
        "  3. Implement missing monitoring and logging controls",
        "",
        "ONGOING:",
        "  1. Maintain quarterly assessment cycle",
        "  2. Update assessments after major network changes",
        "  3. Track remediation progress in Action_Items sheet",
    ]
    
    for line in recommendations_text:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = line
        apply_style(ws[f"A{row}"], styles["exec_summary"])
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    for col in ["D", "E", "F"]:
        ws.column_dimensions[col].width = 20


# ============================================================================
# SECTION 6: SHEET 3 - OVERALL COMPLIANCE
# ============================================================================

def create_overall_compliance_sheet(ws, styles):
    """Create Overall Compliance aggregation sheet."""
    
    ws.title = "Overall_Compliance"
    
    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "Overall Compliance - Aggregate View"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Summary Metrics (referenced by Executive_Summary)
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Key Metrics (Consolidated from WB1-WB5)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    # These formulas link to specific cells in normalized workbooks
    metrics_formulas = [
        ("Total Network Devices", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!B4:B153)"),
        ("Device Hardening Compliance", "=AVERAGE('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Device_Hardening_Assessment'!V4:V153)"),
        ("Network Services Cataloged", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!B4:B103)"),
        ("Security Zones Defined", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Security_Zones_Inventory'!B4:B53)"),
        ("Segmentation Test Pass Rate", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Segmentation_Testing'!G4:G33,\"Pass\")/COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Segmentation_Testing'!G4:G33)*100"),
        ("Average Controls Coverage", "=AVERAGE('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Zone_Control_Assessment'!F4:F33)"),
        ("Total Critical Gaps", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Gap_Summary'!G4:G103,\"Critical\")+COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Gap_Analysis'!F4:F53,\"Critical\")+COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Gap_Analysis'!F4:F53,\"Critical\")+COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Coverage_Gaps'!F4:F43,\"Critical\")"),
    ]
    
    for metric, formula in metrics_formulas:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        
        # Format percentages
        if "Compliance" in metric or "Rate" in metric or "Coverage" in metric:
            ws[f"B{row}"].number_format = "0.0\"%\""
        
        row += 1
    
    # Compliance by Control
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Compliance by ISO 27001:2022 Control"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Compliance Score"
    ws["C" + str(row)] = "Status"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    control_compliance = [
        ("A.8.20 - Networks Security", "=AVERAGE(WB1_Infrastructure!B5,WB2_Device_Security!B5)", "=IF(B" + str(row) + ">=90,\"✔ Compliant\",IF(B" + str(row) + ">=70,\"⚠  Partially Compliant\",\"✗ Non-Compliant\"))"),
        ("A.8.21 - Network Services", "=WB3_Services!B5", "=IF(B" + str(row + 1) + ">=90,\"✔ Compliant\",IF(B" + str(row + 1) + ">=70,\"⚠  Partially Compliant\",\"✗ Non-Compliant\"))"),
        ("A.8.22 - Network Segregation", "=WB4_Segmentation!B5", "=IF(B" + str(row + 2) + ">=90,\"✔ Compliant\",IF(B" + str(row + 2) + ">=70,\"⚠  Partially Compliant\",\"✗ Non-Compliant\"))"),
    ]
    
    for control, score_formula, status_formula in control_compliance:
        ws[f"A{row}"] = control
        ws[f"B{row}"] = score_formula
        ws[f"C{row}"] = status_formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        apply_style(ws[f"C{row}"], styles["metric_value"])
        ws[f"B{row}"].number_format = "0.0\"%\""
        row += 1
    
    # Workbook Summary Links
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Detailed Assessment Links"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Workbook"
    ws["B" + str(row)] = "Navigate To"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    workbook_links = [
        ("WB1: Infrastructure Inventory", "WB1_Infrastructure"),
        ("WB2: Device Security", "WB2_Device_Security"),
        ("WB3: Services Catalog", "WB3_Services"),
        ("WB4: Segmentation Matrix", "WB4_Segmentation"),
        ("WB5: Controls Coverage", "WB5_Controls_Coverage"),
        ("Gap Consolidation", "Gap_Consolidation"),
        ("Action Items", "Action_Items"),
        ("Management Dashboard", "Management_Dashboard"),
    ]
    
    for label, sheet_name in workbook_links:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = f"→ {sheet_name}"
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 3


# ============================================================================
# SECTION 7: SHEET 4 - WB1 INFRASTRUCTURE SUMMARY
# ============================================================================

def create_wb1_infrastructure_sheet(ws, styles):
    """Create WB1 Infrastructure Inventory summary."""
    
    ws.title = "WB1_Infrastructure"
    
    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "WB1: Network Infrastructure Inventory Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Key Metrics
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Infrastructure Inventory Metrics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    wb1_metrics = [
        ("Total Devices Inventoried", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!B4:B153)"),
        ("Critical Devices", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!H4:H153,\"Critical\")"),
        ("High Priority Devices", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!H4:H153,\"High\")"),
        ("Devices Requiring Review", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!Q4:Q153,\"Needs Review\")"),
        ("Discovery Completeness", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!B4:B153)/150*100"),
    ]
    
    for metric, formula in wb1_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        if "Completeness" in metric:
            ws[f"B{row}"].number_format = "0.0\"%\""
        row += 1
    
    # Device Type Breakdown
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Device Type Distribution"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Device Type"
    ws["B" + str(row)] = "Count"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    device_types = ["Router", "Switch", "Firewall", "Wireless AP", "Load Balancer", "VPN Concentrator", "IDS/IPS", "Other"]
    for device_type in device_types:
        ws[f"A{row}"] = device_type
        ws[f"B{row}"] = f"=COUNTIF('[ISMS-IMP-A.8.20-21-22.S1.xlsx]Device_Inventory'!C4:C153,\"{device_type}\")"
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3


# ============================================================================
# SECTION 8: SHEET 5 - WB2 DEVICE SECURITY SUMMARY
# ============================================================================

def create_wb2_device_security_sheet(ws, styles):
    """Create WB2 Device Security Assessment summary."""
    
    ws.title = "WB2_Device_Security"
    
    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "WB2: Device Security Assessment Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Key Metrics
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Device Hardening Compliance Metrics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    wb2_metrics = [
        ("Average Compliance Score", "=AVERAGE('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Device_Hardening_Assessment'!V4:V153)"),
        ("Devices Fully Compliant (≥95%)", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Device_Hardening_Assessment'!V4:V153,\">=95\")"),
        ("Devices Partially Compliant", "=COUNTIFS('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Device_Hardening_Assessment'!V4:V153,\">=80\",'[ISMS-IMP-A.8.20-21-22.S2.xlsx]Device_Hardening_Assessment'!V4:V153,\"<95\")"),
        ("Devices Non-Compliant (<80%)", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Device_Hardening_Assessment'!V4:V153,\"<80\")"),
        ("Total Hardening Gaps", "=SUM('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Device_Hardening_Assessment'!W4:W153)"),
        ("Critical Gaps", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Gap_Summary'!G4:G103,\"Critical\")"),
    ]
    
    for metric, formula in wb2_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        if "Score" in metric:
            ws[f"B{row}"].number_format = "0.0\"%\""
        row += 1
    
    # Top Gaps
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Most Common Hardening Failures"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws[f"A{row}"] = "See Top_Gaps_Analysis sheet in WB2 for detailed breakdown"
    apply_style(ws[f"A{row}"], styles["info_box"])
    
    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3


# ============================================================================
# SECTION 9: SHEET 6 - WB3 SERVICES SUMMARY
# ============================================================================

def create_wb3_services_sheet(ws, styles):
    """Create WB3 Services Catalog summary."""
    
    ws.title = "WB3_Services"
    
    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "WB3: Network Services Security Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Key Metrics
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Network Services Security Metrics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    wb3_metrics = [
        ("Total Services Cataloged", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!B4:B103)"),
        ("Critical Services", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!H4:H103,\"Critical\")"),
        ("Services with Redundancy", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!I4:I103,\"Active-Active\")+COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!I4:I103,\"Active-Passive\")"),
        ("Single Points of Failure", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!I4:I103,\"Single Point of Failure\")"),
        ("Services Monitored", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!J4:J103,\"Monitored\")"),
        ("Service Security Gaps", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Gap_Analysis'!B4:B53)"),
    ]
    
    for metric, formula in wb3_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        row += 1
    
    # Service Type Distribution
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Services by Type"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Service Type"
    ws["B" + str(row)] = "Count"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    service_types = ["DNS", "DHCP", "NTP", "Proxy/Web Filter", "Load Balancer", "RADIUS/TACACS+", "SNMP", "Syslog"]
    for service_type in service_types:
        ws[f"A{row}"] = service_type
        ws[f"B{row}"] = f"=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Services_Catalog'!B4:B103,\"{service_type}\")"
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3


# ============================================================================
# SECTION 10: SHEET 7 - WB4 SEGMENTATION SUMMARY
# ============================================================================

def create_wb4_segmentation_sheet(ws, styles):
    """Create WB4 Segmentation Matrix summary."""
    
    ws.title = "WB4_Segmentation"
    
    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "WB4: Network Segmentation Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Key Metrics
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Network Segmentation Metrics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    wb4_metrics = [
        ("Security Zones Defined", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Security_Zones_Inventory'!B4:B53)"),
        ("VLANs Documented", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]VLAN_Inventory'!A4:A103)"),
        ("Firewall Rules Reviewed", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Firewall_Rules_Assessment'!B4:B103)"),
        ("Segmentation Tests Conducted", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Segmentation_Testing'!B4:B33)"),
        ("Tests Passed", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Segmentation_Testing'!G4:G33,\"Pass\")"),
        ("Tests Failed", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Segmentation_Testing'!G4:G33,\"Fail\")"),
        ("Test Pass Rate", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Segmentation_Testing'!G4:G33,\"Pass\")/COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Segmentation_Testing'!G4:G33)*100"),
        ("Segmentation Gaps", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Gap_Analysis'!B4:B53)"),
    ]
    
    for metric, formula in wb4_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        if "Rate" in metric:
            ws[f"B{row}"].number_format = "0.0\"%\""
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3


# ============================================================================
# SECTION 11: SHEET 8 - WB5 CONTROLS COVERAGE SUMMARY
# ============================================================================

def create_wb5_controls_coverage_sheet(ws, styles):
    """Create WB5 Controls Coverage summary."""
    
    ws.title = "WB5_Controls_Coverage"
    
    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "WB5: Security Controls Coverage Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Key Metrics
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Controls Coverage Metrics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    wb5_metrics = [
        ("Zones Assessed", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Zone_Control_Assessment'!B4:B33)"),
        ("Average Zone Coverage", "=AVERAGE('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Zone_Control_Assessment'!F4:F33)"),
        ("Zones Fully Covered (100%)", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Zone_Control_Assessment'!F4:F33,\"100\")"),
        ("Zones with Coverage Gaps", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Zone_Control_Assessment'!F4:F33,\"<90\")"),
        ("Controls Assessed", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Control_Effectiveness'!A4:A24)"),
        ("Controls Effective", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Control_Effectiveness'!F4:F24,\"Effective\")"),
        ("Coverage Gaps Identified", "=COUNTA('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Coverage_Gaps'!B4:B43)"),
        ("Critical Coverage Gaps", "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Coverage_Gaps'!F4:F43,\"Critical\")"),
    ]
    
    for metric, formula in wb5_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        if "Coverage" in metric and "Average" in metric:
            ws[f"B{row}"].number_format = "0.0\"%\""
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3


# ============================================================================
# SECTION 12: SHEET 9 - GAP CONSOLIDATION
# ============================================================================

def create_gap_consolidation_sheet(ws, styles):
    """Create Gap Consolidation sheet - all gaps from WB1-5."""
    
    ws.title = "Gap_Consolidation"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Gap Consolidation - All Gaps from WB1-WB5"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:F2")
    ws["A2"] = "📋 Consolidated view of all security gaps identified across all 5 assessment workbooks. Prioritize Critical/High gaps first."
    apply_style(ws["A2"], styles["info_box"])
    ws.row_dimensions[2].height = 30
    
    # Gap Summary by Source
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Gap Summary by Source Workbook"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Source"
    ws["B" + str(row)] = "Total Gaps"
    ws["C" + str(row)] = "Critical"
    ws["D" + str(row)] = "High"
    ws["E" + str(row)] = "Medium"
    for col in ["A", "B", "C", "D", "E"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    gap_sources = [
        ("WB2: Device Security", 
         "=COUNTA('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Gap_Summary'!B4:B103)",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Gap_Summary'!G4:G103,\"Critical\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Gap_Summary'!G4:G103,\"High\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S2.xlsx]Gap_Summary'!G4:G103,\"Medium\")"),
        ("WB3: Services Security",
         "=COUNTA('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Gap_Analysis'!B4:B53)",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Gap_Analysis'!F4:F53,\"Critical\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Gap_Analysis'!F4:F53,\"High\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S3.xlsx]Gap_Analysis'!F4:F53,\"Medium\")"),
        ("WB4: Segmentation",
         "=COUNTA('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Gap_Analysis'!B4:B53)",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Gap_Analysis'!F4:F53,\"Critical\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Gap_Analysis'!F4:F53,\"High\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S4.xlsx]Gap_Analysis'!F4:F53,\"Medium\")"),
        ("WB5: Controls Coverage",
         "=COUNTA('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Coverage_Gaps'!B4:B43)",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Coverage_Gaps'!F4:F43,\"Critical\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Coverage_Gaps'!F4:F43,\"High\")",
         "=COUNTIF('[ISMS-IMP-A.8.20-21-22.S5.xlsx]Coverage_Gaps'!F4:F43,\"Medium\")"),
    ]
    
    for source, total, critical, high, medium in gap_sources:
        ws[f"A{row}"] = source
        ws[f"B{row}"] = total
        ws[f"C{row}"] = critical
        ws[f"D{row}"] = high
        ws[f"E{row}"] = medium
        apply_style(ws[f"A{row}"], styles["info_box"])
        for col in ["B", "C", "D", "E"]:
            apply_style(ws[f"{col}{row}"], styles["metric_value"])
        row += 1
    
    # Total row
    ws[f"A{row}"] = "TOTAL"
    ws[f"B{row}"] = f"=SUM(B{row-4}:B{row-1})"
    ws[f"C{row}"] = f"=SUM(C{row-4}:C{row-1})"
    ws[f"D{row}"] = f"=SUM(D{row-4}:D{row-1})"
    ws[f"E{row}"] = f"=SUM(E{row-4}:E{row-1})"
    apply_style(ws[f"A{row}"], styles["column_header"])
    for col in ["B", "C", "D", "E"]:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    
    # Instructions for detailed gaps
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Detailed Gap Information"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    instructions = [
        "For detailed gap information, refer to the Gap Analysis sheets in each workbook:",
        "  • WB2: Device_Hardening_Assessment → Gap_Summary sheet",
        "  • WB3: Services_Catalog → Gap_Analysis sheet",
        "  • WB4: Segmentation_Matrix → Gap_Analysis sheet",
        "  • WB5: Controls_Coverage_Matrix → Coverage_Gaps sheet",
        "",
        "PRIORITIZATION:",
        "  1. CRITICAL gaps: Immediate remediation required",
        "  2. HIGH gaps: Remediate within 30 days",
        "  3. MEDIUM gaps: Remediate within 90 days",
        "  4. LOW gaps: Remediate within 180 days",
    ]
    
    for instruction in instructions:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = instruction
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 3


# ============================================================================
# SECTION 13: SHEET 10 - ACTION ITEMS
# ============================================================================

def create_action_items_sheet(ws, styles):
    """Create Action Items sheet - prioritized remediation roadmap."""
    
    ws.title = "Action_Items"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Action Items - Prioritized Remediation Roadmap"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:F2")
    ws["A2"] = "📋 Prioritized action items for remediation. Update status as gaps are resolved."
    apply_style(ws["A2"], styles["info_box"])
    ws.row_dimensions[2].height = 30
    
    # Priority Actions
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "IMMEDIATE ACTIONS (Critical Priority)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    immediate_actions = [
        "1. Remediate all Critical gaps identified in Gap_Consolidation",
        "2. Implement defense-in-depth for critical zones (see WB5)",
        "3. Address device hardening gaps for critical infrastructure (see WB2)",
        "4. Eliminate single points of failure for critical services (see WB3)",
        "5. Fix failed segmentation tests (see WB4)",
    ]
    
    for action in immediate_actions:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = action
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Short-term actions
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "SHORT-TERM ACTIONS (30-90 Days)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    shortterm_actions = [
        "1. Improve device hardening compliance to ≥95% (see WB2)",
        "2. Implement service security hardening (DNS, DHCP, NTP, Proxy) (see WB3)",
        "3. Complete VLAN segmentation documentation (see WB4)",
        "4. Achieve ≥90% controls coverage for all zones (see WB5)",
        "5. Implement missing monitoring and logging controls",
    ]
    
    for action in shortterm_actions:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = action
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Ongoing actions
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ONGOING ACTIONS"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ongoing_actions = [
        "1. Maintain quarterly network security assessment cycle",
        "2. Update assessments after major network changes",
        "3. Review and update firewall rules quarterly",
        "4. Conduct annual segmentation effectiveness testing",
        "5. Track compliance trends over time",
    ]
    
    for action in ongoing_actions:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = action
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 25


# ============================================================================
# SECTION 14: SHEET 11 - MANAGEMENT DASHBOARD
# ============================================================================

def create_management_dashboard_sheet(ws, styles):
    """Create Management Dashboard - visual executive view."""
    
    ws.title = "Management_Dashboard"
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Network Security Compliance - Management Dashboard"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30
    
    # Key Highlights
    row = 3
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY HIGHLIGHTS"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Assessment Date"
    ws["B" + str(row)] = GENERATED_DATE
    ws["C" + str(row)] = "Overall Status"
    ws["D" + str(row)] = "=IF(Overall_Compliance!B11=0,\"✔ No Critical Gaps\",\"⚠  Critical Gaps Exist\")"
    for col in ["A", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    for col in ["B", "D"]:
        apply_style(ws[col + str(row)], styles["metric_value"])
    
    # Compliance Scorecard
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "COMPLIANCE SCORECARD"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    scorecard = [
        ("A.8.20 - Networks Security", "=Overall_Compliance!B15"),
        ("A.8.21 - Network Services", "=Overall_Compliance!B16"),
        ("A.8.22 - Network Segregation", "=Overall_Compliance!B17"),
        ("Overall Average", "=AVERAGE(B" + str(row + 1) + ":B" + str(row + 3) + ")"),
    ]
    
    for label, formula in scorecard:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        ws[f"B{row}"].number_format = "0.0\"%\""
        row += 1
    
    # Critical Metrics
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CRITICAL METRICS"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    critical_metrics = [
        ("Total Network Devices", "=Overall_Compliance!B5"),
        ("Device Hardening Compliance", "=Overall_Compliance!B6"),
        ("Services with Redundancy", "=WB3_Services!B7"),
        ("Segmentation Test Pass Rate", "=Overall_Compliance!B9"),
        ("Controls Coverage", "=Overall_Compliance!B10"),
        ("Total Critical Gaps", "=Overall_Compliance!B11"),
    ]
    
    for label, formula in critical_metrics:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["metric_label"])
        apply_style(ws[f"B{row}"], styles["metric_value"])
        
        if "Compliance" in label or "Rate" in label or "Coverage" in label:
            ws[f"B{row}"].number_format = "0.0\"%\""
        
        row += 1
    
    # Recommendations
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "EXECUTIVE RECOMMENDATIONS"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    recommendations = [
        "1. Immediate: Remediate all Critical gaps (see Gap_Consolidation)",
        "2. High Priority: Improve device hardening compliance to ≥95%",
        "3. Medium Priority: Enhance service security and redundancy",
        "4. Ongoing: Maintain quarterly assessment cycle",
        "",
        "For detailed findings and action plan, see:",
        "  • Executive_Summary sheet",
        "  • Gap_Consolidation sheet",
        "  • Action_Items sheet",
    ]
    
    for recommendation in recommendations:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = recommendation
        apply_style(ws[f"A{row}"], styles["exec_summary"])
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 15


# ============================================================================
# SECTION 15: MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function - orchestrates dashboard creation."""
    
    logger.info("=" * 80)
    logger.info("ISMS Network Security Compliance Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 - Controls A.8.20, A.8.21, A.8.22")
    logger.info("=" * 80)
    logger.info("")
    
    # Check prerequisites
    if not check_prerequisites():
        return 1
    
    logger.info("─" * 80)
    logger.info("Creating Dashboard Workbook...")
    logger.info("─" * 80)
    
    # Create workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheet_names = [
        "Instructions",
        "Executive_Summary",
        "Overall_Compliance",
        "WB1_Infrastructure",
        "WB2_Device_Security",
        "WB3_Services",
        "WB4_Segmentation",
        "WB5_Controls_Coverage",
        "Gap_Consolidation",
        "Action_Items",
        "Management_Dashboard",
    ]
    
    for name in sheet_names:
        wb.create_sheet(title=name)
    
    # Setup styles
    styles = setup_styles()
    
    # Create all sheets
    logger.info("\n[1/11] Creating Instructions...")
    create_instructions_sheet(wb["Instructions"], styles)
    
    logger.info("[2/11] Creating Executive_Summary...")
    create_executive_summary_sheet(wb["Executive_Summary"], styles)
    
    logger.info("[3/11] Creating Overall_Compliance...")
    create_overall_compliance_sheet(wb["Overall_Compliance"], styles)
    
    logger.info("[4/11] Creating WB1_Infrastructure...")
    create_wb1_infrastructure_sheet(wb["WB1_Infrastructure"], styles)
    
    logger.info("[5/11] Creating WB2_Device_Security...")
    create_wb2_device_security_sheet(wb["WB2_Device_Security"], styles)
    
    logger.info("[6/11] Creating WB3_Services...")
    create_wb3_services_sheet(wb["WB3_Services"], styles)
    
    logger.info("[7/11] Creating WB4_Segmentation...")
    create_wb4_segmentation_sheet(wb["WB4_Segmentation"], styles)
    
    logger.info("[8/11] Creating WB5_Controls_Coverage...")
    create_wb5_controls_coverage_sheet(wb["WB5_Controls_Coverage"], styles)
    
    logger.info("[9/11] Creating Gap_Consolidation...")
    create_gap_consolidation_sheet(wb["Gap_Consolidation"], styles)
    
    logger.info("[10/11] Creating Action_Items...")
    create_action_items_sheet(wb["Action_Items"], styles)
    
    logger.info("[11/11] Creating Management_Dashboard...")
    create_management_dashboard_sheet(wb["Management_Dashboard"], styles)
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.20-21-22.6_Network_Security_Compliance_Dashboard_{GENERATED_TIMESTAMP}.xlsx"
    
    logger.info("\nSaving dashboard...")
    try:
        wb.save(filename)
        logger.info(f"✅ SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"❌ ERROR saving workbook: {e}")
        return 1
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("✅ DASHBOARD GENERATED SUCCESSFULLY")
    logger.info("=" * 80)
    logger.info(f"\nDashboard File: {filename}")
    logger.info(f"Generated: {GENERATED_DATE}")
    logger.info(f"Sheets: {len(sheet_names)}")
    logger.info("\nData Sources (External Links):")
    for wb_file in REQUIRED_WORKBOOKS:
        logger.info(f"  ✅ {wb_file}")
    logger.info("\n" + "=" * 80)
    logger.info("🎯 NEXT STEPS:")
    logger.info("  1. Open the dashboard file")
    logger.info("  2. Review Executive_Summary for high-level findings")
    logger.info("  3. Check Management_Dashboard for visual overview")
    logger.info("  4. Review Gap_Consolidation for prioritized gaps")
    logger.info("  5. Track remediation in Action_Items sheet")
    logger.info("\n💡 IMPORTANT:")
    logger.info("  • Keep all workbooks (WB1-WB5) in same directory")
    logger.info("  • Dashboard auto-updates when source workbooks change")
    logger.info("  • Use 'Data → Edit Links' to manage external links")
    logger.info("\n" + "=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
