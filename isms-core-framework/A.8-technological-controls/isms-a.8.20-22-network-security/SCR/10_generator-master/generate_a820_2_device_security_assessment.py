#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-21-22.S2 - Network Device Security Assessment Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 2 of 6: Network Device Hardening and Security Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific hardening baselines, security standards, and
assessment requirements.

Key customization areas:
1. Hardening baseline requirements (match your security standards)
2. Device-specific security controls (vendor and device type specific)
3. Compliance scoring criteria (adapt to your risk profile)
4. Authentication requirements (based on your AAA infrastructure)
5. Gap severity classification (aligned with your risk assessment)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-21-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
network device security configurations against hardening baselines, supporting
evidence-based validation of device security controls.

**Purpose:**
Enables comprehensive security assessment of network devices against hardening
baselines, supporting evidence-based validation of device security configurations
and identification of gaps in compliance with ISO 27001:2022 Control A.8.20.

**Assessment Scope:**
- Device hardening baseline compliance
- Authentication mechanisms (local, AAA, 802.1X)
- Access control configuration (privilege levels, command authorization)
- Encryption configuration (management protocols, wireless encryption)
- Logging and monitoring configuration
- Configuration backup status
- Patch/update status
- Gap identification and remediation tracking

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and hardening standards
2. Device_Security_Summary - Overall device security posture dashboard
3. Router_Hardening - Router-specific security assessment
4. Switch_Hardening - Switch-specific security assessment
5. Firewall_Hardening - Firewall-specific security assessment
6. Wireless_Hardening - Wireless AP-specific security assessment
7. LoadBalancer_Hardening - Load balancer-specific security assessment
8. VPN_Hardening - VPN concentrator-specific security assessment
9. Authentication_Controls - Authentication and access control assessment
10. Encryption_Controls - Encryption configuration assessment
11. Logging_Monitoring - Logging and monitoring configuration assessment
12. Gap_Analysis - Security gaps and remediation tracking
13. Evidence_Register - Audit evidence tracking and documentation
14. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with security control dropdown lists
- Conditional formatting for compliance status (compliant/non-compliant/partial)
- Automated compliance scoring per device
- Protected formulas with unprotected input cells
- Gap severity classification (Critical/High/Medium/Low)
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with configuration management systems

**Integration:**
This assessment imports device inventory from WB1 (Infrastructure Inventory)
and feeds security posture data into WB5 (Controls Coverage) and the Network
Security Compliance Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_2_device_security_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_2_device_security_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_2_device_security_assessment.py --date 20250124
    
    # Generate with custom organization name
    python3 generate_a820_2_device_security_assessment.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organization name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_2_Device_Security_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize hardening baselines to match your security standards
    2. Import device list from WB1 (Infrastructure Inventory)
    3. Assess each device against applicable hardening baseline
    4. Document security control implementation status
    5. Validate configurations against vendor hardening guides (CIS, vendor docs)
    6. Identify gaps and classify by severity
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (config exports, scan results)
    9. Obtain stakeholder approvals
    10. Feed results into WB5 and Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    2 of 6 (Network Device Hardening and Security Controls)
Primary Control:      A.8.20 (Network Security)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-21-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-21-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-21-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-21-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-21-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-21-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-21-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-21-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-21-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-21-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-21-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-21-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalization)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-21-22 specification
    - Supports comprehensive network device security assessment
    - Integrated with A.8.20-21-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Hardening Baseline Standards:**
Device hardening baselines should reference industry standards:
- CIS Benchmarks (Center for Internet Security)
- Vendor hardening guides (Cisco, Juniper, Palo Alto, etc.)
- DISA STIGs (for government/defense sectors)
- NIST SP 800-53 controls
- Organization-specific security standards

Review and update baselines quarterly to incorporate new security requirements.

**Technology Diversity:**
This assessment framework is technology-agnostic and must work across:
- Traditional network devices (physical infrastructure)
- Virtual network appliances (NFV, virtual firewalls)
- Software-Defined Networks (SDN controllers, OpenFlow switches)
- Cloud-based network services (AWS, Azure, GCP)

Customize assessment criteria to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Documented hardening baselines for each device type
- Evidence of configuration compliance (config exports, scan results)
- Gap analysis with remediation plans
- Timelines and ownership for remediation activities
- Re-assessment results post-remediation

**Data Protection:**
Assessment workbooks contain sensitive security information including:
- Device security configurations and vulnerabilities
- Authentication and access control details
- Identified security gaps and weaknesses
- Remediation plans revealing security posture

Handle in accordance with your organization's data classification policies.
Restrict access to authorized security personnel only.

**Maintenance:**
Review and update security assessments:
- Monthly: After device configuration changes or patches
- Quarterly: Routine reassessment of critical devices
- Annually: Complete reassessment of all devices
- Ad-hoc: After security incidents or vulnerability disclosures

**Quality Assurance:**
Validate assessment accuracy by:
- Exporting actual device configurations (show run, get config)
- Running automated compliance scanners (Nessus, Qualys, etc.)
- Manual verification of critical security controls
- Peer review by network security engineers
- Testing security controls (penetration testing)

**Regulatory Alignment:**
Ensure hardening standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 network device hardening requirements
- Healthcare: HIPAA network security configuration standards
- Finance: Regional banking network security requirements
- Government: Jurisdiction-specific hardening mandates

Customize assessment criteria to include regulatory-specific requirements.

**Integration Points:**
This assessment integrates with other ISO 27001 controls:
- A.8.8 (Vulnerability Management): Device vulnerability scanning results
- A.8.9 (Configuration Management): Configuration baseline compliance
- A.8.15 (Logging): Device logging configuration assessment
- A.8.16 (Monitoring): Device monitoring configuration assessment

**Common Pitfalls to Avoid:**
1. **Generic Baselines**: Using one-size-fits-all hardening (customize per device!)
2. **No Evidence**: Configuration claims without actual config exports
3. **Checkbox Compliance**: Marking "compliant" without verification
4. **Ignoring Legacy**: Excluding old devices from assessment
5. **No Remediation**: Identifying gaps without action plans
6. **Default Credentials**: Missing default password changes
7. **Unnecessary Services**: Not disabling unused protocols (Telnet, HTTP)
8. **Weak Encryption**: Allowing SSLv3, TLS 1.0, or weak ciphers

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, Reference
except ImportError as e:
    logger.error(f"❌ ERROR: Required library not found: {e}")
    logger.info("📦 Install required libraries: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================

WORKBOOK_NAME = "Network Device Security Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.20-21-22.S2"
CONTROL_REF = "ISO/IEC 27001:2022 - Controls A.8.20, A.8.21, A.8.22: Network Security"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Network_Device_Security_Assessment_{GENERATED_TIMESTAMP}.xlsx"

# Assessment constants
DEVICE_ROW_COUNT = 150      # Device assessment rows
GAP_ROW_COUNT = 100         # Gap tracking rows
REMEDIATION_ROW_COUNT = 50  # Remediation roadmap rows

# Hardening requirements (checklist items)
HARDENING_REQUIREMENTS = [
    "Default Credentials Disabled",
    "Strong Password Policy",
    "Multi-Factor Authentication (MFA)",
    "Unnecessary Services Disabled",
    "Secure Management (SSH/HTTPS Only)",
    "Logging Enabled",
    "NTP Configured",
    "SNMPv3 Only (v1/v2c Disabled)",
    "Session Timeouts Configured",
    "Banner Messages Configured",
    "Configuration Backups",
    "Firmware/Software Up-to-Date",
    "Access Control Lists (ACLs)",
    "Encrypted Management Traffic",
    "Least Privilege Access",
]

# Number of hardening requirements
NUM_REQUIREMENTS = len(HARDENING_REQUIREMENTS)


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "yes_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="006100"),
        },
        "no_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="9C0006"),
        },
        "na_fill": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "font": Font(name="Calibri", size=10, color="7F7F7F"),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
        },
        "low_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
        },
        "compliant_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "noncompliant_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "partial_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "info_box": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Yes/No/N/A validation (for hardening requirements)
    validations["yes_no_na"] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False,
    )
    validations["yes_no_na"].error = "Must be Yes, No, or N/A"
    validations["yes_no_na"].errorTitle = "Invalid Value"
    
    # Device Type validation
    validations["device_type"] = DataValidation(
        type="list",
        formula1='"Router,Switch,Firewall,Wireless AP,Load Balancer,VPN Concentrator,IDS/IPS,Network Management,Other"',
        allow_blank=False,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚪ Low"',
        allow_blank=False,
    )
    
    # Remediation Status validation
    validations["remediation_status"] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Completed,Accepted Risk,Deferred"',
        allow_blank=False,
    )
    
    # Compliance Status validation
    validations["compliance_status"] = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant,Partially Compliant,Not Assessed"',
        allow_blank=False,
    )
    
    return validations


# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & ASSESSMENT GUIDE
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Assessment Guide sheet."""
    
    ws.title = "Instructions & Guide"
    
    # Title with Document ID and ISO Control Reference
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 40
    
    # Document Information
    ws.merge_cells("A3:B3")
    ws["A3"] = "Document Information"
    apply_style(ws["A3"], styles["header"])
    
    info_data = [
        ("Workbook:", WORKBOOK_NAME),
        ("Generated:", GENERATED_DATE),
        ("Version:", "1.0"),
        ("Control:", "ISO 27001:2022 A.8.20 (Networks Security)"),
        ("Purpose:", "Assess device hardening compliance against security baselines"),
        ("Related IMP:", "ISMS-IMP-A.8.20-21-22-S3 (Device Hardening Process)"),
        ("Prerequisite:", "Workbook 1 (Device Inventory) should be completed"),
    ]
    
    row = 4
    for label, value in info_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Assessment Methodology
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Assessment Methodology"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    methodology = [
        "1. PREPARATION: Review Device_Inventory from Workbook 1, identify devices to assess",
        "2. BASELINE REVIEW: Study Hardening_Baseline_Reference sheet (CIS Benchmarks, vendor guides)",
        "3. DEVICE ASSESSMENT: For each device, evaluate compliance with each hardening requirement",
        "4. SCORING: Use Yes/No/N/A for each requirement. Compliance score auto-calculates.",
        "5. GAP IDENTIFICATION: Document non-compliant items in Gap_Summary sheet",
        "6. REMEDIATION: Prioritize gaps by severity, create remediation plan",
        "7. VALIDATION: Re-assess after remediation, update compliance scores",
        "8. EVIDENCE: Maintain configuration backups, scan results as evidence",
    ]
    
    for instruction in methodology:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = instruction
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Assessment Criteria
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Assessment Response Criteria"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Response"
    ws["B" + str(row)] = "Definition"
    ws["C" + str(row)] = "When to Use"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    criteria = [
        ("Yes", "Requirement is fully implemented and verified", "Hardening control is configured correctly and confirmed"),
        ("No", "Requirement is NOT implemented or incorrectly configured", "Control is missing, disabled, or misconfigured"),
        ("N/A", "Requirement does not apply to this device", "Device type/role does not require this control"),
    ]
    
    row += 1
    for response, definition, when_to_use in criteria:
        ws[f"A{row}"] = response
        ws[f"B{row}"] = definition
        ws[f"C{row}"] = when_to_use
        
        if response == "Yes":
            apply_style(ws[f"A{row}"], styles["yes_fill"])
        elif response == "No":
            apply_style(ws[f"A{row}"], styles["no_fill"])
        elif response == "N/A":
            apply_style(ws[f"A{row}"], styles["na_fill"])
        
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Compliance Scoring
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Compliance Scoring Formula"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = 'Compliance Score (%) = (Count of "Yes") / (Count of "Yes" + Count of "No") × 100'
    apply_style(ws[f"A{row}"], styles["info_box"])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = '📌 NOTE: "N/A" responses are excluded from scoring (not applicable controls do not affect compliance)'
    apply_style(ws[f"A{row}"], styles["info_box"])
    
    # Hardening Requirements Overview
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = f"Hardening Requirements Overview ({NUM_REQUIREMENTS} Requirements)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "#"
    ws["B" + str(row)] = "Requirement"
    ws["C" + str(row)] = "Applicability"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    applicability_notes = [
        "All devices",
        "All devices",
        "Management devices, critical infrastructure",
        "All devices (disable unused services)",
        "All devices (SSH for CLI, HTTPS for web)",
        "All devices (syslog, local logs)",
        "All devices (time synchronization)",
        "All devices (if SNMP used)",
        "All devices (idle timeout)",
        "All devices (login banner)",
        "All devices (automated backups)",
        "All devices (security patches)",
        "Routers, Firewalls (ingress/egress filtering)",
        "All devices (TLS for management)",
        "All devices (role-based access)",
    ]
    
    row += 1
    for idx, req in enumerate(HARDENING_REQUIREMENTS, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = req
        ws[f"C{row}"] = applicability_notes[idx - 1] if idx <= len(applicability_notes) else "All devices"
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Important Notes
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "⚠ IMPORTANT ASSESSMENT NOTES"
    apply_style(ws[f"A{row}"], styles["header"])
    
    notes = [
        "• Base assessments on CIS Benchmarks and vendor hardening guides (see Hardening_Baseline_Reference)",
        "• Verify configurations via CLI, GUI, or config file review (don't rely on documentation alone)",
        "• Use 'N/A' appropriately - don't mark everything N/A to inflate scores",
        "• Document evidence for each assessment (config screenshots, command outputs)",
        "• Reassess after remediation to track improvement",
        "• Critical/High severity gaps require immediate remediation",
    ]
    
    row += 1
    for note in notes:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = note
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15


# ============================================================================
# SECTION 5: SHEET 2 - DEVICE HARDENING ASSESSMENT (MAIN)
# ============================================================================

def create_device_hardening_assessment_sheet(ws, styles, validations):
    """Create main Device Hardening Assessment sheet."""
    
    ws.title = "Device_Hardening_Assessment"
    
    # Title
    title_col_span = 5 + NUM_REQUIREMENTS + 3  # Device info + requirements + scoring
    ws.merge_cells(f"A1:{get_column_letter(title_col_span)}1")
    cell = ws["A1"]
    cell.value = f"Network Device Hardening Assessment - Generated {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells(f"A2:{get_column_letter(title_col_span)}2")
    ws["A2"] = f"📋 Assess each device against {NUM_REQUIREMENTS} hardening requirements. Yellow cells are assessment fields (Yes/No/N/A). Compliance score auto-calculates."
    apply_style(ws["A2"], styles["info_box"])
    ws.row_dimensions[2].height = 30
    
    # Column Headers
    headers = [
        "Device ID",           # A (from WB1)
        "Device Type",         # B
        "Hostname",            # C
        "Primary IP",          # D
        "Criticality",         # E
    ]
    
    # Add hardening requirement columns
    for req in HARDENING_REQUIREMENTS:
        headers.append(req)
    
    # Add scoring columns
    headers.extend([
        "Compliance Score (%)",  # Auto-calculated
        "Gap Count",             # Auto-calculated
        "Assessment Date",
        "Assessor",
        "Notes",
    ])
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx <= 5 else 12
    
    # Data rows
    start_row = 4
    end_row = start_row + DEVICE_ROW_COUNT - 1
    
    req_start_col = 6  # Requirements start at column F
    req_end_col = 5 + NUM_REQUIREMENTS
    score_col = req_end_col + 1
    gap_col = score_col + 1
    
    for row_idx in range(start_row, end_row + 1):
        # Device info columns (A-E) - input cells
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Hardening requirement columns (F onwards) - Yes/No/N/A dropdowns
        for col in range(req_start_col, req_end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Compliance Score formula
        # Count "Yes" / (Count "Yes" + Count "No") * 100
        yes_range = f"{get_column_letter(req_start_col)}{row_idx}:{get_column_letter(req_end_col)}{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=score_col, value=formula)
        ws.cell(row=row_idx, column=score_col).number_format = "0.0"
        
        # Gap Count formula (count "No" responses)
        gap_formula = f'=COUNTIF({yes_range},"No")'
        ws.cell(row=row_idx, column=gap_col, value=gap_formula)
        
        # Assessment Date, Assessor, Notes
        for col in range(gap_col + 1, gap_col + 4):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations to hardening requirement columns
    for col in range(req_start_col, req_end_col + 1):
        validations["yes_no_na"].add(
            f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}"
        )
    ws.add_data_validation(validations["yes_no_na"])
    
    # Conditional Formatting for Yes/No/N/A columns
    from openpyxl.formatting.rule import CellIsRule
    
    for col in range(req_start_col, req_end_col + 1):
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        
        # Yes = Green
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        # No = Red
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
        # N/A = Gray
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"N/A"'], fill=styles["na_fill"]["fill"], font=styles["na_fill"]["font"])
        )
    
    # Conditional Formatting for Compliance Score
    score_col_letter = get_column_letter(score_col)
    score_range = f"{score_col_letter}{start_row}:{score_col_letter}{end_row}"
    
    # >= 95% = Green (Compliant)
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["95"], fill=styles["compliant_fill"]["fill"])
    )
    # 80-94% = Yellow (Partially Compliant)
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="between", formula=["80", "94"], fill=styles["partial_fill"]["fill"])
    )
    # < 80% = Red (Non-Compliant)
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="lessThan", formula=["80"], fill=styles["noncompliant_fill"]["fill"])
    )
    
    # Freeze panes
    ws.freeze_panes = "F4"
    
    # Column widths for scoring columns
    ws.column_dimensions[get_column_letter(score_col)].width = 15
    ws.column_dimensions[get_column_letter(gap_col)].width = 10
    ws.column_dimensions[get_column_letter(gap_col + 1)].width = 15
    ws.column_dimensions[get_column_letter(gap_col + 2)].width = 20
    ws.column_dimensions[get_column_letter(gap_col + 3)].width = 40


# ============================================================================
# SECTION 6: SHEET 3 - HARDENING BASELINE REFERENCE
# ============================================================================

def create_hardening_baseline_reference_sheet(ws, styles):
    """Create Hardening Baseline Reference with CIS Benchmarks."""
    
    ws.title = "Hardening_Baseline_Reference"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Device Hardening Baseline Reference"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Purpose
    ws.merge_cells("A2:F2")
    ws["A2"] = "Reference guide for device hardening based on CIS Benchmarks and vendor security guides"
    apply_style(ws["A2"], styles["info_box"])
    
    # CIS Benchmarks
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CIS Benchmarks & Industry Standards"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Device Type"
    ws["B" + str(row)] = "Benchmark/Guide"
    ws["C" + str(row)] = "Version"
    ws["D" + str(row)] = "URL/Reference"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    benchmarks = [
        ("Cisco IOS", "CIS Cisco IOS Benchmark", "v4.1.0", "https://www.cisecurity.org/benchmark/cisco"),
        ("Juniper JunOS", "CIS Juniper Benchmark", "v2.1.0", "https://www.cisecurity.org/benchmark/juniper"),
        ("Palo Alto", "CIS Palo Alto Firewall Benchmark", "v1.1.0", "https://www.cisecurity.org/benchmark/palo_alto_networks"),
        ("Fortinet", "CIS Fortinet FortiOS Benchmark", "v1.1.0", "https://www.cisecurity.org/benchmark/fortinet"),
        ("Aruba", "Aruba Hardening Guide", "Latest", "https://www.arubanetworks.com/techdocs/"),
        ("Generic Network", "NIST SP 800-41 (Firewall/Router)", "Rev 1", "https://csrc.nist.gov/publications/detail/sp/800-41/rev-1/final"),
        ("Wireless", "CIS Wireless Network Security", "v1.0", "https://www.cisecurity.org/"),
    ]
    
    row += 1
    for device, benchmark, version, url in benchmarks:
        ws[f"A{row}"] = device
        ws[f"B{row}"] = benchmark
        ws[f"C{row}"] = version
        ws[f"D{row}"] = url
        for col in ["A", "B", "C", "D"]:
            apply_style(ws[f"{col}{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Hardening Requirements Detail
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Hardening Requirements - Implementation Guidance"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Requirement"
    ws["B" + str(row)] = "Implementation Guidance"
    ws["C" + str(row)] = "Verification Method"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    guidance = [
        (
            "Default Credentials Disabled",
            "Change default usernames/passwords (admin/admin, cisco/cisco, etc.). Disable or rename default accounts.",
            "Attempt login with default credentials (should fail). Review user list."
        ),
        (
            "Strong Password Policy",
            "Min 12 chars, complexity (upper/lower/number/special), expiration 90 days, history 5 passwords.",
            "Check password policy config. Test weak password (should be rejected)."
        ),
        (
            "Multi-Factor Authentication",
            "Enable MFA for admin access (TACACS+, RADIUS with OTP/tokens). Critical for privileged access.",
            "Verify AAA config, test login (should require 2nd factor)."
        ),
        (
            "Unnecessary Services Disabled",
            "Disable unused protocols: HTTP (use HTTPS), Telnet (use SSH), FTP, TFTP, finger, echo, CDP (on edge).",
            "Port scan device, check running services list."
        ),
        (
            "Secure Management (SSH/HTTPS)",
            "Use SSH for CLI (disable Telnet), HTTPS for web (disable HTTP). SSH v2 only.",
            "Telnet to device (should fail). Check SSH version."
        ),
        (
            "Logging Enabled",
            "Enable syslog to centralized server. Log severity: informational or higher. Buffer logging.",
            "Check logging config, verify logs reaching syslog server."
        ),
        (
            "NTP Configured",
            "Configure NTP client, point to internal NTP servers. Enable NTP authentication if possible.",
            "Check NTP config, verify time sync (ntpstat, show ntp status)."
        ),
        (
            "SNMPv3 Only",
            "Use SNMPv3 with authentication and encryption. Disable SNMPv1/v2c (community strings).",
            "Check SNMP config, test v2c query (should fail)."
        ),
        (
            "Session Timeouts",
            "Configure idle timeout (e.g., 10 minutes for console, 15 min for VTY). Auto-logout inactive sessions.",
            "Check timeout config (exec-timeout, idle-timeout)."
        ),
        (
            "Banner Messages",
            "Configure login banner warning unauthorized access is prohibited. MOTD and login banners.",
            "Check banner config, verify displayed on login."
        ),
        (
            "Configuration Backups",
            "Automated nightly config backups to secure location. Use RANCID, Oxidized, or scripts.",
            "Verify backup schedule, check recent backup files exist."
        ),
        (
            "Firmware/Software Updated",
            "Running latest stable firmware/software with security patches. Track end-of-life dates.",
            "Check current version vs. vendor latest. Review security advisories."
        ),
        (
            "Access Control Lists",
            "Implement ACLs on routers/firewalls for ingress/egress filtering. Default deny, explicit allow.",
            "Review ACL config, test traffic (permitted/denied as expected)."
        ),
        (
            "Encrypted Management",
            "Encrypt management traffic: SSH (not Telnet), HTTPS (not HTTP), SNMPv3, TLS syslog.",
            "Capture management traffic (tcpdump), verify encrypted."
        ),
        (
            "Least Privilege Access",
            "Role-based access control (RBAC). Users have minimum privileges needed. Privilege levels configured.",
            "Review user privilege levels, test restricted user access."
        ),
    ]
    
    row += 1
    for requirement, impl, verify in guidance:
        ws[f"A{row}"] = requirement
        ws[f"B{row}"] = impl
        ws[f"C{row}"] = verify
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 40
        row += 1
    
    # Device-Specific Notes
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Device-Specific Hardening Notes"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    device_notes = [
        ("Routers", "Disable IP source routing, enable uRPF, implement ACLs, secure routing protocols (MD5 auth)."),
        ("Switches", "Enable port security, DHCP snooping, DAI, disable DTP, change native VLAN, disable unused ports."),
        ("Firewalls", "Default deny policy, rule cleanup, logging all denied traffic, regular rule review."),
        ("Wireless APs", "WPA3 or WPA2-Enterprise, disable WPS, hide SSID (optional), MAC filtering, rogue AP detection."),
        ("Load Balancers", "Strong SSL/TLS ciphers, disable SSLv3/TLS1.0, enable HSTS, session persistence security."),
    ]
    
    for device, notes in device_notes:
        ws[f"A{row}"] = device
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = notes
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


# ============================================================================
# SECTION 7: SHEET 4 - GAP SUMMARY
# ============================================================================

def create_gap_summary_sheet(ws, styles, validations):
    """Create Gap Summary sheet for identified hardening gaps."""
    
    ws.title = "Gap_Summary"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "Device Hardening Gaps - Summary & Remediation Tracking"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = '📋 Document all "No" responses from Device_Hardening_Assessment. Track remediation progress and priority.'
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",                # A - Auto
        "Device ID",             # B
        "Device Type",           # C
        "Hostname",              # D
        "Hardening Requirement", # E (which requirement failed)
        "Current State",         # F (description)
        "Gap Severity",          # G (Critical/High/Medium/Low)
        "Remediation Plan",      # H
        "Owner",                 # I
        "Status",                # J (dropdown)
        "Target Date",           # K
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Gap ID (auto-generated)
        gap_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","HARD-GAP-" & TEXT({gap_num},"000"),"")')
        
        # Input cells
        for col in range(2, 12):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["gap_severity"].add(f"G{start_row}:G{end_row}")
    ws.add_data_validation(validations["gap_severity"])
    
    validations["remediation_status"].add(f"J{start_row}:J{end_row}")
    ws.add_data_validation(validations["remediation_status"])
    
    # Conditional Formatting for Severity
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Gap Severity Guidelines
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "Gap Severity Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Severity"
    ws["B" + str(row)] = "Definition"
    ws["C" + str(row)] = "Examples"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    severity_guide = [
        ("Critical", "Immediate exploitation risk, default credentials, no encryption", "Default passwords still enabled, Telnet enabled on Internet-facing device"),
        ("High", "Significant risk, missing key controls", "No MFA for admin access, SNMPv1/v2c enabled, no logging"),
        ("Medium", "Moderate risk, important controls missing", "Weak password policy, no session timeouts, outdated firmware (non-critical)"),
        ("Low", "Minor gap, limited risk", "Missing login banner, unused ports not disabled"),
    ]
    
    row += 1
    for severity, defn, examples in severity_guide:
        ws[f"A{row}"] = severity
        ws[f"B{row}"] = defn
        ws[f"C{row}"] = examples
        
        if severity == "Critical":
            apply_style(ws[f"A{row}"], styles["critical_fill"])
        elif severity == "High":
            apply_style(ws[f"A{row}"], styles["high_fill"])
        elif severity == "Medium":
            apply_style(ws[f"A{row}"], styles["medium_fill"])
        elif severity == "Low":
            apply_style(ws[f"A{row}"], styles["low_fill"])
        
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 12
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SHEET 5 - COMPLIANCE SCORING
# ============================================================================

def create_compliance_scoring_sheet(ws, styles):
    """Create Compliance Scoring sheet with overall metrics."""
    
    ws.title = "Compliance_Scoring"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Overall Hardening Compliance Scoring"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Summary Statistics
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Summary Statistics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    ws["C" + str(row)] = "Formula"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    # Calculate column reference for compliance score in main sheet
    req_start = 6
    req_end = 5 + NUM_REQUIREMENTS
    score_col = req_end + 1
    score_col_letter = get_column_letter(score_col)
    
    stats = [
        ("Total Devices Assessed", f'=COUNTA(Device_Hardening_Assessment!A4:A{3 + DEVICE_ROW_COUNT})'),
        ("Devices with Score > 0%", f'=COUNTIF(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},">0")'),
        ("Average Compliance Score", f'=AVERAGE(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT})'),
        ("Compliant Devices (≥95%)", f'=COUNTIF(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},">=95")'),
        ("Partially Compliant (80-94%)", f'=COUNTIFS(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},">=80",Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},"<95")'),
        ("Non-Compliant (<80%)", f'=COUNTIF(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},"<80")'),
        ("Total Identified Gaps", f'=SUM(Device_Hardening_Assessment!{get_column_letter(score_col + 1)}4:{get_column_letter(score_col + 1)}{3 + DEVICE_ROW_COUNT})'),
    ]
    
    row += 1
    for metric, formula in stats:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        ws[f"C{row}"] = formula  # Show formula
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        
        # Format percentage for average
        if "Average" in metric:
            ws[f"B{row}"].number_format = "0.0"
        
        row += 1
    
    # Compliance Status Distribution
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Compliance Status Distribution"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Status"
    ws["B" + str(row)] = "Count"
    ws["C" + str(row)] = "Percentage"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    chart_start_row = row + 1
    row += 1
    
    statuses = [
        ("Compliant (≥95%)", f'=COUNTIF(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},">=95")'),
        ("Partially Compliant (80-94%)", f'=COUNTIFS(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},">=80",Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},"<95")'),
        ("Non-Compliant (<80%)", f'=COUNTIF(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},"<80")'),
        ("Not Assessed", f'=COUNTIF(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},"")'),
    ]
    
    for status, formula in statuses:
        ws[f"A{row}"] = status
        ws[f"B{row}"] = formula
        ws[f"C{row}"] = f"=B{row}/B5*100"  # Percentage (B5 has total devices)
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws[f"C{row}"].number_format = "0.0%"
        row += 1
    
    chart_end_row = row - 1
    
    # Create Pie Chart - Compliance Distribution
    pie_chart = PieChart()
    pie_chart.title = "Compliance Status Distribution"
    pie_chart.style = 10
    labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_end_row)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.height = 10
    pie_chart.width = 15
    ws.add_chart(pie_chart, f"D{chart_start_row}")
    
    # Compliance Targets
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Compliance Targets & Status"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Target"
    ws["C" + str(row)] = "Current"
    ws["D" + str(row)] = "Status"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    targets = [
        ("Average Compliance Score", "≥ 95%", f"=B{5}", f'=IF(C{row + 1}>=95,"✔ Met","✗ Not Met")'),
        ("Compliant Devices %", "≥ 80%", f"=B{8}/B{5}*100", f'=IF(C{row + 2}>=80,"✔ Met","✗ Not Met")'),
        ("Critical Gaps", "= 0", f"=COUNTIF(Gap_Summary!G4:G{3 + GAP_ROW_COUNT},\"Critical\")", f'=IF(C{row + 3}=0,"✔ Met","✗ Not Met")'),
        ("High Gaps Remediation", "≥ 80%", f"=COUNTIFS(Gap_Summary!G4:G{3 + GAP_ROW_COUNT},\"High\",Gap_Summary!J4:J{3 + GAP_ROW_COUNT},\"Completed\")/COUNTIF(Gap_Summary!G4:G{3 + GAP_ROW_COUNT},\"High\")*100", f'=IF(C{row + 4}>=80,"✔ Met","✗ Not Met")'),
    ]
    
    row += 1
    for metric, target, current, status in targets:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = target
        ws[f"C{row}"] = current
        ws[f"D{row}"] = status
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        
        if "%" in metric:
            ws[f"C{row}"].number_format = "0.0"
        
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 3


# ============================================================================
# SECTION 9: SHEET 6 - DEVICE TYPE COMPLIANCE
# ============================================================================

def create_device_type_compliance_sheet(ws, styles):
    """Create Device Type Compliance breakdown."""
    
    ws.title = "Device_Type_Compliance"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Compliance Analysis by Device Type"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Purpose
    ws.merge_cells("A2:F2")
    ws["A2"] = "Breakdown of hardening compliance scores by device type to identify problem areas"
    apply_style(ws["A2"], styles["info_box"])
    
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Average Compliance Score by Device Type"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Device Type"
    ws["B" + str(row)] = "Device Count"
    ws["C" + str(row)] = "Avg Compliance Score"
    ws["D" + str(row)] = "Min Score"
    ws["E" + str(row)] = "Max Score"
    for col in ["A", "B", "C", "D", "E"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    device_types = [
        "Router",
        "Switch",
        "Firewall",
        "Wireless AP",
        "Load Balancer",
        "VPN Concentrator",
        "IDS/IPS",
        "Network Management",
        "Other",
    ]
    
    # Get score column from main sheet
    req_end = 5 + NUM_REQUIREMENTS
    score_col = req_end + 1
    score_col_letter = get_column_letter(score_col)
    
    chart_start_row = row + 1
    row += 1
    
    for dev_type in device_types:
        ws[f"A{row}"] = dev_type
        ws[f"B{row}"] = f'=COUNTIF(Device_Hardening_Assessment!B4:B{3 + DEVICE_ROW_COUNT},"{dev_type}")'
        ws[f"C{row}"] = f'=AVERAGEIF(Device_Hardening_Assessment!B4:B{3 + DEVICE_ROW_COUNT},"{dev_type}",Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT})'
        ws[f"D{row}"] = f'=MINIFS(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},Device_Hardening_Assessment!B4:B{3 + DEVICE_ROW_COUNT},"{dev_type}")'
        ws[f"E{row}"] = f'=MAXIFS(Device_Hardening_Assessment!{score_col_letter}4:{score_col_letter}{3 + DEVICE_ROW_COUNT},Device_Hardening_Assessment!B4:B{3 + DEVICE_ROW_COUNT},"{dev_type}")'
        
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        apply_style(ws[f"E{row}"], styles["info_box"])
        
        ws[f"C{row}"].number_format = "0.0"
        ws[f"D{row}"].number_format = "0.0"
        ws[f"E{row}"].number_format = "0.0"
        
        row += 1
    
    chart_end_row = row - 1
    
    # Create Bar Chart - Compliance by Device Type
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Average Compliance Score by Device Type"
    bar_chart.y_axis.title = "Compliance Score (%)"
    bar_chart.x_axis.title = "Device Type"
    labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    data = Reference(ws, min_col=3, min_row=chart_start_row, max_row=chart_end_row)
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.height = 12
    bar_chart.width = 18
    ws.add_chart(bar_chart, f"A{row + 2}")
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 3


# ============================================================================
# SECTION 10: SHEET 7 - TOP GAPS ANALYSIS
# ============================================================================

def create_top_gaps_analysis_sheet(ws, styles):
    """Create Top Gaps Analysis showing most common failures."""
    
    ws.title = "Top_Gaps_Analysis"
    
    # Title
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "Top Hardening Gaps - Most Common Failures"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Purpose
    ws.merge_cells("A2:E2")
    ws["A2"] = "Identify most frequently failed hardening requirements across all devices"
    apply_style(ws["A2"], styles["info_box"])
    
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"Hardening Requirements Failure Analysis ({NUM_REQUIREMENTS} Requirements)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Requirement"
    ws["B" + str(row)] = "Total 'No' Count"
    ws["C" + str(row)] = "Failure Rate %"
    ws["D" + str(row)] = "Priority"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    # Calculate failure counts for each requirement
    req_start_col = 6
    chart_start_row = row + 1
    row += 1
    
    for idx, req in enumerate(HARDENING_REQUIREMENTS):
        col_letter = get_column_letter(req_start_col + idx)
        ws[f"A{row}"] = req
        ws[f"B{row}"] = f'=COUNTIF(Device_Hardening_Assessment!{col_letter}4:{col_letter}{3 + DEVICE_ROW_COUNT},"No")'
        ws[f"C{row}"] = f'=B{row}/COUNTA(Device_Hardening_Assessment!{col_letter}4:{col_letter}{3 + DEVICE_ROW_COUNT})*100'
        ws[f"D{row}"] = f'=IF(C{row}>=75,"CRITICAL",IF(C{row}>=50,"HIGH",IF(C{row}>=25,"MEDIUM","LOW")))'
        
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        
        ws[f"C{row}"].number_format = "0.0%"
        
        row += 1
    
    chart_end_row = row - 1
    
    # Create Bar Chart - Top Gaps
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Most Common Hardening Failures"
    bar_chart.y_axis.title = "Failure Count"
    bar_chart.x_axis.title = "Requirement"
    labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_end_row)
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.height = 15
    bar_chart.width = 20
    ws.add_chart(bar_chart, f"A{row + 2}")
    
    # Conditional Formatting for Priority
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"HIGH"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"LOW"'], fill=styles["low_fill"]["fill"])
    )
    
    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 3


# ============================================================================
# SECTION 11: SHEET 8 - REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap_sheet(ws, styles, validations):
    """Create Remediation Roadmap with prioritized action plan."""
    
    ws.title = "Remediation_Roadmap"
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Hardening Remediation Roadmap"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "📋 Prioritized remediation plan sorted by severity. Focus on Critical/High gaps first."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Priority",            # A (1, 2, 3...)
        "Severity",            # B (Critical/High/Medium/Low)
        "Gap Description",     # C
        "Affected Devices",    # D (count or list)
        "Remediation Action",  # E
        "Owner",               # F
        "Target Date",         # G
        "Status",              # H (dropdown)
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + REMEDIATION_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Priority (manual numbering)
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        
        # Input cells
        for col in range(2, 9):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["gap_severity"].add(f"B{start_row}:B{end_row}")
    ws.add_data_validation(validations["gap_severity"])
    
    validations["remediation_status"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["remediation_status"])
    
    # Conditional Formatting for Severity
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Remediation Approach
    row = end_row + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Remediation Approach Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    approach = [
        "1. PRIORITIZE BY SEVERITY: Address Critical gaps immediately, then High, then Medium/Low",
        "2. CLUSTER BY REQUIREMENT: Remediate same requirement across multiple devices together (efficiency)",
        "3. LOW-HANGING FRUIT: Quick wins (e.g., enable logging, configure banners) → build momentum",
        "4. COMPLEX ITEMS: MFA, major config changes require planning, testing, approval",
        "5. VALIDATE: Test in lab/dev environment before production deployment",
        "6. DOCUMENT: Maintain change records, configuration backups, validation evidence",
        "7. RE-ASSESS: After remediation, re-run assessment to confirm compliance improvement",
    ]
    
    for item in approach:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = item
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 12: MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.20-21-22 - Network Device Security Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.20 (Networks Security) - Workbook 2")
    logger.info("=" * 80)
    logger.info("\n🎯 Systems Engineering Approach: Systematic Hardening Assessment")
    logger.info("📊 CIS Benchmarks & Vendor Guides: Industry-standard baselines")
    logger.info("🔒 Audit-Ready: Compliance scoring and gap tracking")
    logger.info("\n" + "─" * 80)
    
    # Create workbook
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets
    sheet_names = [
        "Instructions & Guide",
        "Device_Hardening_Assessment",
        "Hardening_Baseline_Reference",
        "Gap_Summary",
        "Compliance_Scoring",
        "Device_Type_Compliance",
        "Top_Gaps_Analysis",
        "Remediation_Roadmap",
    ]
    
    for name in sheet_names:
        wb.create_sheet(title=name)
    
    logger.info(f"✅ Workbook created with {len(sheet_names)} sheets")
    
    # Setup styles and validations
    logger.info("\n[Phase 2] Setting up styles and data validations...")
    styles = setup_styles()
    validations = create_data_validations()
    logger.info("✅ Styles and validations configured")
    
    # Create all sheets
    logger.info("\n[Phase 3] Generating assessment sheets...")
    
    logger.info("  [1/8] Creating Instructions & Guide...")
    create_instructions_sheet(wb["Instructions & Guide"], styles)
    logger.info(f"  ✅ Instructions complete (methodology, {NUM_REQUIREMENTS} requirements overview)")
    
    logger.info("  [2/8] Creating Device_Hardening_Assessment...")
    create_device_hardening_assessment_sheet(wb["Device_Hardening_Assessment"], styles, validations)
    logger.info(f"  ✅ Main assessment complete ({DEVICE_ROW_COUNT} device rows, {NUM_REQUIREMENTS} hardening requirements)")
    
    logger.info("  [3/8] Creating Hardening_Baseline_Reference...")
    create_hardening_baseline_reference_sheet(wb["Hardening_Baseline_Reference"], styles)
    logger.info("  ✅ Baseline reference complete (CIS Benchmarks, vendor guides, implementation guidance)")
    
    logger.info("  [4/8] Creating Gap_Summary...")
    create_gap_summary_sheet(wb["Gap_Summary"], styles, validations)
    logger.info(f"  ✅ Gap summary complete ({GAP_ROW_COUNT} gap tracking rows)")
    
    logger.info("  [5/8] Creating Compliance_Scoring...")
    create_compliance_scoring_sheet(wb["Compliance_Scoring"], styles)
    logger.info("  ✅ Compliance scoring complete (metrics, targets, pie chart)")
    
    logger.info("  [6/8] Creating Device_Type_Compliance...")
    create_device_type_compliance_sheet(wb["Device_Type_Compliance"], styles)
    logger.info("  ✅ Device type compliance complete (breakdown by device type, bar chart)")
    
    logger.info("  [7/8] Creating Top_Gaps_Analysis...")
    create_top_gaps_analysis_sheet(wb["Top_Gaps_Analysis"], styles)
    logger.error(f"  ✅ Top gaps analysis complete (failure analysis, {NUM_REQUIREMENTS} requirements)")
    
    logger.info("  [8/8] Creating Remediation_Roadmap...")
    create_remediation_roadmap_sheet(wb["Remediation_Roadmap"], styles, validations)
    logger.info(f"  ✅ Remediation roadmap complete ({REMEDIATION_ROW_COUNT} action items)")
    
    # Save workbook
    logger.info("\n[Phase 4] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.20-21-22.S2_Network_Device_Security_Assessment_{GENERATED_TIMESTAMP}.xlsx"
    
    try:
        wb.save(filename)
        logger.info(f"✅ SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"❌ ERROR saving workbook: {e}")
        return 1
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("📋 WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\n📊 Assessment Sheets:")
    logger.info("  • Instructions & Guide (methodology, assessment criteria)")
    logger.info(f"  • Device_Hardening_Assessment ({DEVICE_ROW_COUNT} devices × {NUM_REQUIREMENTS} requirements)")
    logger.info("  • Hardening_Baseline_Reference (CIS Benchmarks, vendor guides)")
    logger.info(f"  • Gap_Summary ({GAP_ROW_COUNT} gap tracking rows)")
    logger.info("\n📈 Analysis & Reporting:")
    logger.info("  • Compliance_Scoring (overall metrics, targets, pie chart)")
    logger.info("  • Device_Type_Compliance (breakdown by device type, bar chart)")
    logger.error("  • Top_Gaps_Analysis (most common failures, bar chart)")
    logger.info(f"  • Remediation_Roadmap ({REMEDIATION_ROW_COUNT} prioritized actions)")
    logger.info("\n" + "─" * 80)
    logger.info("📈 ASSESSMENT CAPABILITIES:")
    logger.info(f"  • {DEVICE_ROW_COUNT} device hardening assessments")
    logger.info(f"  • {NUM_REQUIREMENTS} hardening requirements per device")
    logger.info("  • Auto-calculated compliance scores (Yes/No/N/A → %)")
    logger.info("  • Conditional formatting (Yes=Green, No=Red, N/A=Gray)")
    logger.info(f"  • {GAP_ROW_COUNT} gap identification and tracking")
    logger.info(f"  • {REMEDIATION_ROW_COUNT} remediation roadmap items")
    logger.info("  • 3 charts (compliance pie, device type bar, top gaps bar)")
    logger.info("  • CIS Benchmark references and implementation guidance")
    logger.info("\n" + "─" * 80)
    logger.info("🎯 KEY FEATURES:")
    logger.info("  ✅ Based on CIS Benchmarks and vendor hardening guides")
    logger.info("  ✅ Systematic Yes/No/N/A assessment per requirement")
    logger.info("  ✅ Automated compliance scoring formulas")
    logger.info("  ✅ Gap identification with severity classification")
    logger.info("  ✅ Remediation tracking and prioritization")
    logger.info("  ✅ Device type compliance analysis")
    logger.info("  ✅ Top gaps visualization")
    logger.info("  ✅ Integration with Workbook 1 (Device Inventory)")
    logger.info("\n" + "=" * 80)
    logger.info("🚀 NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Read Instructions & Guide sheet first")
    logger.info("  3. Review Hardening_Baseline_Reference (CIS Benchmarks)")
    logger.info("  4. For each device from WB1, assess hardening compliance")
    logger.info("  5. Fill Device_Hardening_Assessment (Yes/No/N/A per requirement)")
    logger.info("  6. Document all 'No' responses in Gap_Summary")
    logger.info("  7. Review Compliance_Scoring and Device_Type_Compliance")
    logger.info("  8. Prioritize remediation using Remediation_Roadmap")
    logger.info("  9. Re-assess after remediation to track improvement")
    logger.info("\n💡 PRO TIP:")
    logger.info("  Start with Critical devices first (from WB1 criticality assessment).")
    logger.info("  Use Hardening_Baseline_Reference as your assessment guide.")
    logger.info("  Don't mark everything 'N/A' - be honest about gaps to improve security.")
    logger.info("\n" + "=" * 80)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info('— and you are the easiest person to fool." - Richard Feynman')
    logger.info("\n🎁 This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
