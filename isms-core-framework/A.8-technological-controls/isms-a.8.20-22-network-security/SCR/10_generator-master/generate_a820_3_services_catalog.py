#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-21-22.S3 - Network Services Catalog Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 3 of 6: Network Services Inventory and Security Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific network services catalog, security standards, and
assessment requirements.

Key customization areas:
1. Network services inventory (match your actual service catalog)
2. Service-specific security requirements (based on your service architecture)
3. Availability requirements and SLAs (aligned with business needs)
4. Redundancy and failover criteria (based on your resilience requirements)
5. Monitoring and alerting thresholds (adapted to your operations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-21-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for inventorying
and assessing network services security across the organization, supporting
evidence-based validation of service security controls.

**Purpose:**
Enables systematic inventory and security assessment of network services (DNS,
DHCP, NTP, proxy, load balancers, etc.), supporting evidence-based validation
of service security controls against ISO 27001:2022 Control A.8.21.

**Assessment Scope:**
- Network services inventory (DNS, DHCP, NTP, proxy, LB, etc.)
- Service-specific security assessment tabs
- Service availability and redundancy status
- Security mechanisms per service type
- Service monitoring and alerting configuration
- Compliance scoring per service
- Service owner/responsible party tracking
- Gap identification and remediation planning

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and service security standards
2. Services_Summary - Overall network services security dashboard
3. DNS_Services - DNS security assessment (DNSSEC, split DNS, etc.)
4. DHCP_Services - DHCP security assessment (rogue prevention, etc.)
5. NTP_Services - NTP security assessment (time source authentication)
6. Proxy_Services - Proxy security assessment (filtering, authentication)
7. LoadBalancer_Services - Load balancer security (SSL termination, etc.)
8. Authentication_Services - Auth services (RADIUS, TACACS+, LDAP)
9. SNMP_Services - SNMP security (v3 only, community string management)
10. Syslog_Services - Syslog security (encryption, authentication)
11. Other_Services - Additional network services
12. Availability_Monitoring - Service availability and uptime tracking
13. Gap_Analysis - Service security gaps and remediation tracking
14. Evidence_Register - Audit evidence tracking and documentation
15. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with service type and security mechanism dropdown lists
- Conditional formatting for service security status
- Automated compliance scoring per service
- Protected formulas with unprotected input cells
- Service availability SLA tracking
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with service monitoring systems

**Integration:**
This assessment works alongside WB1 (Infrastructure Inventory) to document
services running on network devices, and feeds into WB5 (Controls Coverage)
and the Network Security Compliance Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_3_services_catalog.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_3_services_catalog.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_3_services_catalog.py --date 20250124
    
    # Generate with custom organization name
    python3 generate_a820_3_services_catalog.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organization name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_3_Services_Catalog_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize service types to match your service catalog
    2. Inventory all network services across infrastructure
    3. Complete service-specific security assessments
    4. Document security mechanisms for each service
    5. Validate service availability and redundancy configurations
    6. Review monitoring and alerting setup for each service
    7. Identify gaps and classify by severity
    8. Define remediation actions with timelines and ownership
    9. Collect and link audit evidence (service configs, logs)
    10. Obtain stakeholder approvals
    11. Feed results into WB5 and Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    3 of 6 (Network Services Inventory and Security Assessment)
Primary Control:      A.8.21 (Security of Network Services)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-21-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-21-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-21-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-21-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-21-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-21-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-21-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-21-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-21-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-21-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-21-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-21-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalization)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-21-22 specification
    - Supports comprehensive network services security assessment
    - Integrated with A.8.20-21-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Network Services Security Standards:**
Network services security should follow industry best practices:
- DNS: DNSSEC implementation, split DNS architecture, query logging
- DHCP: Rogue DHCP prevention, DHCP snooping, scope management
- NTP: Stratum hierarchy, time source authentication, symmetric keys
- Proxy: Content filtering, authentication, SSL inspection, logging
- Load Balancers: SSL/TLS termination, session management, health checks

Review and update service security requirements quarterly.

**Technology Diversity:**
This assessment framework is technology-agnostic and must work across:
- Traditional on-premise services (physical/virtual servers)
- Cloud-based services (AWS Route53, Azure DNS, managed services)
- Hybrid architectures (split DNS, hybrid proxy, etc.)
- Containerized services (Kubernetes services, service mesh)

Customize assessment criteria to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Complete network services inventory
- Service-specific security controls documented
- Evidence of security mechanism implementation
- Availability monitoring and SLA compliance data
- Gap analysis with remediation plans

**Data Protection:**
Assessment workbooks contain sensitive service information including:
- Service architecture and topology
- Security mechanisms and configurations
- Authentication and access control details
- Identified vulnerabilities and gaps

Handle in accordance with your organization's data classification policies.
Restrict access to authorized network and security personnel only.

**Maintenance:**
Review and update services assessment:
- Monthly: After service configuration changes or new services deployed
- Quarterly: Routine reassessment of critical services
- Annually: Complete reassessment of all network services
- Ad-hoc: After service outages or security incidents

**Quality Assurance:**
Validate service security by:
- Exporting service configurations (zone files, DHCP scopes, etc.)
- Testing security mechanisms (DNSSEC validation, DHCP snooping tests)
- Reviewing service logs and monitoring data
- Peer review by service owners and security engineers
- Penetration testing of service security controls

**Regulatory Alignment:**
Ensure network services security aligns with applicable regulatory requirements:
- Payment processing: PCI DSS network services requirements
- Healthcare: HIPAA network services security standards
- Finance: Regional banking network services requirements
- Government: Jurisdiction-specific service security mandates

Customize assessment criteria to include regulatory-specific requirements.

**Integration Points:**
This assessment integrates with other ISO 27001 controls:
- A.8.15 (Logging): Service logging requirements and implementation
- A.8.16 (Monitoring): Service monitoring and availability tracking
- A.8.20 (Network Security): Services running on network devices
- A.5.23 (Cloud Services): Cloud-based network services

**Common Pitfalls to Avoid:**
1. **DNS Vulnerabilities**: Cache poisoning, zone transfers, no DNSSEC
2. **DHCP Attacks**: Rogue DHCP servers, DHCP exhaustion, no snooping
3. **NTP Attacks**: Amplification attacks, unauthenticated time sources
4. **Proxy Bypass**: Poor filtering rules, SSL/TLS inspection gaps
5. **Load Balancer Issues**: Weak SSL/TLS, session hijacking risks
6. **SNMP v1/v2**: Using insecure SNMP versions with community strings
7. **Unencrypted Syslog**: Logs transmitted without encryption
8. **No Redundancy**: Single points of failure for critical services

**Service-Specific Security Considerations:**

**DNS Security:**
- Implement DNSSEC for zone signing and validation
- Use split DNS (internal vs. external resolution)
- Disable zone transfers to unauthorized hosts
- Enable query logging and monitoring
- Protect against DNS tunneling and exfiltration

**DHCP Security:**
- Enable DHCP snooping on switches
- Implement rogue DHCP server detection
- Use DHCP reservations for critical systems
- Log DHCP lease assignments
- Protect against DHCP exhaustion attacks

**NTP Security:**
- Use NTP version 4 with authentication
- Implement time source hierarchy (stratum levels)
- Use symmetric keys for authentication
- Restrict NTP access with ACLs
- Monitor time synchronization status

**Proxy Security:**
- Implement SSL/TLS inspection where appropriate
- Use content filtering and malware scanning
- Enforce authentication for proxy access
- Log all proxy traffic
- Implement bypass prevention

**Load Balancer Security:**
- Use secure SSL/TLS configurations (TLS 1.2+)
- Implement proper session management
- Configure health checks for backend servers
- Enable logging and monitoring
- Protect against DDoS attacks

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
except ImportError as e:
    logger.error(f"❌ ERROR: Required library not found: {e}")
    logger.info("📦 Install required libraries: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================

WORKBOOK_NAME = "Network Services Catalog & Security Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.20-21-22.S3"
CONTROL_REF = "ISO/IEC 27001:2022 - Controls A.8.20, A.8.21, A.8.22: Network Security"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Network_Services_Catalog_{GENERATED_TIMESTAMP}.xlsx"

# Service catalog constants
SERVICE_ROW_COUNT = 100         # Main service catalog rows
DNS_ASSESSMENT_ROWS = 20        # DNS services
DHCP_ASSESSMENT_ROWS = 15       # DHCP services
NTP_ASSESSMENT_ROWS = 10        # NTP services
PROXY_ASSESSMENT_ROWS = 15      # Proxy services
ADDITIONAL_SERVICE_ROWS = 30    # Other services (LB, AAA, SNMP, Syslog)
GAP_ROW_COUNT = 50              # Gap tracking

# Service types
SERVICE_TYPES = [
    "DNS",
    "DHCP",
    "NTP",
    "Proxy/Web Filter",
    "Load Balancer",
    "RADIUS/TACACS+ (AAA)",
    "SNMP",
    "Syslog",
    "SMTP Relay",
    "FTP/SFTP",
    "Other",
]


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "yes_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="006100"),
        },
        "no_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="9C0006"),
        },
        "na_fill": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "font": Font(name="Calibri", size=10, color="7F7F7F"),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        },
        "low_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "compliant_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "noncompliant_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "partial_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "info_box": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Service Type validation
    validations["service_type"] = DataValidation(
        type="list",
        formula1=f'"{",".join(SERVICE_TYPES)}"',
        allow_blank=False,
    )
    validations["service_type"].error = "Invalid service type"
    validations["service_type"].errorTitle = "Service Type Error"
    
    # Criticality validation
    validations["criticality"] = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚪ Low"',
        allow_blank=False,
    )
    
    # Status validation
    validations["status"] = DataValidation(
        type="list",
        formula1='"Active,Offline,Planned,Decommissioned"',
        allow_blank=False,
    )
    
    # Yes/No/N/A validation
    validations["yes_no_na"] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False,
    )
    
    # Redundancy Level validation
    validations["redundancy"] = DataValidation(
        type="list",
        formula1='"Active-Active,Active-Passive,Clustered,Single Point of Failure,N/A"',
        allow_blank=True,
    )
    
    # Monitoring Status validation
    validations["monitoring"] = DataValidation(
        type="list",
        formula1='"Monitored,Not Monitored,Partially Monitored"',
        allow_blank=True,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚪ Low"',
        allow_blank=False,
    )
    
    # Gap Status validation
    validations["gap_status"] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Accepted Risk"',
        allow_blank=False,
    )
    
    return validations


# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & SERVICE GUIDE
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Service Guide sheet."""
    
    ws.title = "Instructions & Guide"
    
    # Title with Document ID and ISO Control Reference
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 40
    
    # Document Information
    ws.merge_cells("A3:B3")
    ws["A3"] = "Document Information"
    apply_style(ws["A3"], styles["header"])
    
    info_data = [
        ("Workbook:", WORKBOOK_NAME),
        ("Generated:", GENERATED_DATE),
        ("Version:", "1.0"),
        ("Control:", "ISO 27001:2022 A.8.21 (Security of Network Services)"),
        ("Purpose:", "Catalog network services and assess security controls"),
        ("Related IMP:", "ISMS-IMP-A.8.20-21-22-S4 (Services Security Process)"),
    ]
    
    row = 4
    for label, value in info_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Assessment Approach
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Assessment Approach"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    approach = [
        "1. SERVICE DISCOVERY: Identify all network services (DNS, DHCP, NTP, Proxy, Load Balancer, AAA, SNMP, Syslog)",
        "2. CATALOG: Document each service in Services_Catalog (purpose, hosting, criticality, redundancy)",
        "3. SERVICE-SPECIFIC ASSESSMENT: Complete dedicated assessment sheets (DNS, DHCP, NTP, Proxy)",
        "4. SECURITY EVALUATION: Assess security controls per service type (authentication, encryption, logging, etc.)",
        "5. AVAILABILITY: Document redundancy, failover, SLA requirements",
        "6. MONITORING: Verify service health monitoring and alerting",
        "7. GAP IDENTIFICATION: Document missing controls or misconfigurations",
        "8. REMEDIATION: Prioritize and track service security improvements",
    ]
    
    for instruction in approach:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = instruction
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Service Type Definitions
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Network Service Type Definitions"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Service Type"
    ws["B" + str(row)] = "Purpose"
    ws["C" + str(row)] = "Key Security Controls"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    service_defs = [
        ("DNS", "Domain name resolution", "DNSSEC, Split DNS, Rate Limiting, Query Logging"),
        ("DHCP", "IP address assignment", "DHCP Snooping, Rogue Detection, Scope Management"),
        ("NTP", "Time synchronization", "Authentication, Access Control, Stratum Hierarchy"),
        ("Proxy/Web Filter", "Web traffic filtering", "Authentication, SSL Inspection, Logging, Bypass Prevention"),
        ("Load Balancer", "Traffic distribution", "SSL Termination, Health Checks, Session Persistence"),
        ("RADIUS/TACACS+", "AAA services", "Strong Authentication, Command Authorization, Logging"),
        ("SNMP", "Device monitoring", "SNMPv3 Only, Authentication, Encryption"),
        ("Syslog", "Centralized logging", "TLS Encryption, Retention, Rotation"),
        ("SMTP Relay", "Email routing", "Authentication, Relay Controls, SPF/DKIM"),
        ("FTP/SFTP", "File transfer", "Encryption (SFTP), Authentication, Access Control"),
    ]
    
    row += 1
    for service, purpose, controls in service_defs:
        ws[f"A{row}"] = service
        ws[f"B{row}"] = purpose
        ws[f"C{row}"] = controls
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Service Criticality Guidelines
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Service Criticality Assessment Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Criticality"
    ws["B" + str(row)] = "Definition"
    ws["C" + str(row)] = "Examples"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    crit_guide = [
        ("Critical", "Service failure causes immediate business impact", "Internal DNS, Primary DHCP, Authentication (RADIUS/TACACS+)"),
        ("High", "Service failure causes significant degradation", "Web Proxy, External DNS, Load Balancer"),
        ("Medium", "Service failure causes localized impact", "NTP, Secondary DHCP, SNMP monitoring"),
        ("Low", "Service failure causes minimal impact", "Guest Wi-Fi DHCP, Development services"),
    ]
    
    row += 1
    for crit, defn, examples in crit_guide:
        ws[f"A{row}"] = crit
        ws[f"B{row}"] = defn
        ws[f"C{row}"] = examples
        
        if crit == "Critical":
            apply_style(ws[f"A{row}"], styles["critical_fill"])
        elif crit == "High":
            apply_style(ws[f"A{row}"], styles["high_fill"])
        elif crit == "Medium":
            apply_style(ws[f"A{row}"], styles["medium_fill"])
        elif crit == "Low":
            apply_style(ws[f"A{row}"], styles["low_fill"])
        
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Important Notes
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "⚠ IMPORTANT ASSESSMENT NOTES"
    apply_style(ws[f"A{row}"], styles["header"])
    
    notes = [
        "• Each service type has dedicated assessment sheet with specific security controls",
        "• Services_Catalog is the master list - link to detailed assessments",
        "• Verify redundancy for all Critical/High services (no single points of failure)",
        "• Ensure all services are monitored (health checks, alerting)",
        "• Document service dependencies (e.g., DNS required for Active Directory)",
        "• Review service hardening per IMP-S4 guidance",
    ]
    
    row += 1
    for note in notes:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = note
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    for col in ["D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 15


# ============================================================================
# SECTION 5: SHEET 2 - SERVICES CATALOG (MAIN)
# ============================================================================

def create_services_catalog_sheet(ws, styles, validations):
    """Create main Services Catalog sheet."""
    
    ws.title = "Services_Catalog"
    
    # Title
    ws.merge_cells("A1:P1")
    cell = ws["A1"]
    cell.value = f"Network Services Catalog - Generated {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "📋 Master catalog of all network services. Yellow cells are input fields. Link to detailed assessments in service-specific sheets."
    apply_style(ws["A2"], styles["info_box"])
    ws.row_dimensions[2].height = 30
    
    # Column Headers
    headers = [
        "Service ID",           # A - Auto
        "Service Type",         # B - Dropdown
        "Service Name",         # C
        "Purpose",              # D
        "Hosting Location",     # E (Device/Server)
        "IP Address",           # F
        "Port(s)",              # G
        "Criticality",          # H - Dropdown
        "Redundancy Level",     # I - Dropdown
        "Monitoring Status",    # J - Dropdown
        "SLA Uptime %",         # K
        "Owner",                # L
        "Last Security Review", # M
        "Status",               # N - Dropdown
        "Notes",                # O
        "Assessment Link",      # P (reference to detailed sheet)
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + SERVICE_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Service ID (auto-generated)
        service_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SVC-" & TEXT({service_num},"000"),"")')
        
        # Input cells
        for col in range(2, 17):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["service_type"].add(f"B{start_row}:B{end_row}")
    ws.add_data_validation(validations["service_type"])
    
    validations["criticality"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["criticality"])
    
    validations["redundancy"].add(f"I{start_row}:I{end_row}")
    ws.add_data_validation(validations["redundancy"])
    
    validations["monitoring"].add(f"J{start_row}:J{end_row}")
    ws.add_data_validation(validations["monitoring"])
    
    validations["status"].add(f"N{start_row}:N{end_row}")
    ws.add_data_validation(validations["status"])
    
    # Conditional Formatting for Criticality
    
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Column widths
    widths = [12, 18, 25, 35, 20, 15, 10, 12, 20, 15, 12, 20, 15, 12, 40, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: SHEET 3 - DNS SECURITY ASSESSMENT
# ============================================================================

def create_dns_security_sheet(ws, styles, validations):
    """Create DNS Security Assessment sheet."""
    
    ws.title = "DNS_Security_Assessment"
    
    # Title
    ws.merge_cells("A1:N1")
    cell = ws["A1"]
    cell.value = "DNS Security Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:N2")
    ws["A2"] = "📋 Assess DNS security controls per IMP-S4 guidance. One row per DNS service instance."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers - DNS-specific security controls
    headers = [
        "Service ID",           # A (from catalog)
        "DNS Type",             # B (Authoritative/Recursive/Forwarder)
        "DNSSEC Enabled",       # C - Y/N/NA
        "Split DNS",            # D - Y/N/NA
        "Rate Limiting",        # E - Y/N/NA
        "Query Logging",        # F - Y/N/NA
        "RPZ/Filtering",        # G - Y/N/NA
        "Recursion Control",    # H - Y/N/NA
        "Cache Poisoning Protection", # I - Y/N/NA
        "Zone Transfer Controls", # J - Y/N/NA
        "TSIG/SIG(0)",          # K - Y/N/NA
        "Compliance Score %",   # L - Auto-calc
        "Last Assessed",        # M
        "Notes",                # N
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + DNS_ASSESSMENT_ROWS - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Service ID
        cell = ws.cell(row=row_idx, column=1)
        apply_style(cell, styles["input_cell"])
        
        # DNS Type dropdown
        cell = ws.cell(row=row_idx, column=2)
        apply_style(cell, styles["input_cell"])
        
        # Security controls (Yes/No/N/A)
        for col in range(3, 12):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Compliance Score formula
        yes_range = f"C{row_idx}:K{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=12, value=formula)
        ws.cell(row=row_idx, column=12).number_format = "0.0"
        
        # Last Assessed, Notes
        for col in [13, 14]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Yes/No/NA validation to control columns
    for col in range(3, 12):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])
    
    # Apply conditional formatting for Yes/No/NA
    for col in range(3, 12):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"N/A"'], fill=styles["na_fill"]["fill"], font=styles["na_fill"]["font"])
        )
    
    # DNS Security Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "DNS Security Controls - Implementation Reference (per IMP-S4)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    dns_controls = [
        ("DNSSEC", "Enable zone signing, publish DS records to parent, configure trust anchors"),
        ("Split DNS", "Internal view for private zones, external view for public zones"),
        ("Rate Limiting", "Configure response rate limiting (RRL) to prevent DDoS amplification"),
        ("Query Logging", "Enable query logging for security monitoring and forensics"),
        ("RPZ/Filtering", "Implement Response Policy Zones for malware/phishing domain blocking"),
        ("Recursion Control", "Disable recursion on authoritative servers, restrict recursive queries by ACL"),
        ("Cache Poisoning Protection", "Enable DNSSEC validation, source port randomization"),
        ("Zone Transfer Controls", "Restrict zone transfers (AXFR) to authorized secondary servers only"),
        ("TSIG/SIG(0)", "Use TSIG for authenticated zone transfers and dynamic updates"),
    ]
    
    row += 1
    for control, guidance in dns_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:N{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["N"].width = 40


# ============================================================================
# SECTION 7: SHEET 4 - DHCP SECURITY ASSESSMENT
# ============================================================================

def create_dhcp_security_sheet(ws, styles, validations):
    """Create DHCP Security Assessment sheet."""
    
    ws.title = "DHCP_Security_Assessment"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "DHCP Security Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "📋 Assess DHCP security controls per IMP-S4 guidance. One row per DHCP server/scope."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "DHCP Scope",           # B (IP range)
        "DHCP Snooping",        # C - Y/N/NA
        "Rogue DHCP Detection", # D - Y/N/NA
        "Server Hardening",     # E - Y/N/NA
        "Lease Time Appropriate", # F - Y/N/NA
        "Utilization Monitoring", # G - Y/N/NA
        "Reservation Management", # H - Y/N/NA
        "Logging Enabled",      # I - Y/N/NA
        "Compliance Score %",   # J - Auto
        "Last Assessed",        # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + DHCP_ASSESSMENT_ROWS - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Service ID, DHCP Scope
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Security controls
        for col in range(3, 10):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Compliance Score
        yes_range = f"C{row_idx}:I{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=10, value=formula)
        ws.cell(row=row_idx, column=10).number_format = "0.0"
        
        # Last Assessed, Notes
        for col in [11, 12]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply validations and formatting
    for col in range(3, 10):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])
    
    for col in range(3, 10):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
    
    # DHCP Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "DHCP Security Controls - Implementation Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    dhcp_controls = [
        ("DHCP Snooping", "Enable on access switches to prevent rogue DHCP servers"),
        ("Rogue Detection", "Active scanning or monitoring for unauthorized DHCP responses"),
        ("Server Hardening", "Apply OS hardening, disable unnecessary services, patch regularly"),
        ("Lease Time", "Appropriate lease time (8 hours workstations, 24 hours servers, 1 hour guest)"),
        ("Utilization Monitoring", "Monitor scope utilization, alert at 80% threshold"),
        ("Reservation Management", "Document static reservations, review regularly"),
        ("Logging", "Enable DHCP server logging for all lease actions"),
    ]
    
    row += 1
    for control, guidance in dhcp_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:L{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["L"].width = 40


# ============================================================================
# SECTION 8: SHEET 5 - NTP SECURITY ASSESSMENT
# ============================================================================

def create_ntp_security_sheet(ws, styles, validations):
    """Create NTP Security Assessment sheet."""
    
    ws.title = "NTP_Security_Assessment"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "NTP Security Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "📋 Assess NTP security controls per IMP-S4 guidance. One row per NTP server/client."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "NTP Role",             # B (Server/Client)
        "Authentication",       # C - Y/N/NA
        "Access Control",       # D - Y/N/NA
        "Stratum Appropriate",  # E - Y/N/NA
        "Source Validation",    # F - Y/N/NA
        "Monitoring",           # G - Y/N/NA
        "Compliance Score %",   # H - Auto
        "Stratum Level",        # I
        "Last Assessed",        # J
        "Notes",                # K
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + NTP_ASSESSMENT_ROWS - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Service ID, NTP Role
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Security controls
        for col in range(3, 8):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Compliance Score
        yes_range = f"C{row_idx}:G{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=8, value=formula)
        ws.cell(row=row_idx, column=8).number_format = "0.0"
        
        # Stratum, Last Assessed, Notes
        for col in [9, 10, 11]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply validations
    for col in range(3, 8):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])
    
    for col in range(3, 8):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
    
    # NTP Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "NTP Security Controls - Implementation Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    ntp_controls = [
        ("Authentication", "Enable NTP authentication (MD5 keys) between servers"),
        ("Access Control", "Restrict NTP queries and modifications (restrict default kod nomodify notrap)"),
        ("Stratum Hierarchy", "Appropriate stratum levels (Stratum 1/2 for internal servers)"),
        ("Source Validation", "Verify NTP server sources, use multiple reliable sources"),
        ("Monitoring", "Monitor time offset, alert on significant drift (>1 second)"),
    ]
    
    row += 1
    for control, guidance in ntp_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["K"].width = 40


# ============================================================================
# SECTION 9: SHEET 6 - PROXY SECURITY ASSESSMENT
# ============================================================================

def create_proxy_security_sheet(ws, styles, validations):
    """Create Proxy/Web Filter Security Assessment sheet."""
    
    ws.title = "Proxy_Security_Assessment"
    
    # Title
    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "Proxy/Web Filter Security Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "📋 Assess Proxy/Web Filtering security controls per IMP-S4 guidance."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "Proxy Type",           # B (Forward/Reverse/Transparent)
        "Authentication",       # C - Y/N/NA
        "SSL Inspection",       # D - Y/N/NA
        "Category Filtering",   # E - Y/N/NA
        "Malware Scanning",     # F - Y/N/NA
        "Logging",              # G - Y/N/NA
        "Bypass Prevention",    # H - Y/N/NA
        "Redundancy",           # I - Y/N/NA
        "Compliance Score %",   # J - Auto
        "Last Assessed",        # K
        "Filter Categories",    # L
        "Notes",                # M
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + PROXY_ASSESSMENT_ROWS - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Service ID, Proxy Type
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Security controls
        for col in range(3, 10):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Compliance Score
        yes_range = f"C{row_idx}:I{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=10, value=formula)
        ws.cell(row=row_idx, column=10).number_format = "0.0"
        
        # Last Assessed, Filter Categories, Notes
        for col in [11, 12, 13]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply validations
    for col in range(3, 10):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])
    
    for col in range(3, 10):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
    
    # Proxy Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "Proxy Security Controls - Implementation Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    proxy_controls = [
        ("Authentication", "LDAP/AD authentication, Kerberos/NTLM for transparent proxy"),
        ("SSL Inspection", "SSL/TLS decryption and inspection (with CA cert deployment)"),
        ("Category Filtering", "Block malicious/inappropriate categories (malware, adult, gambling, etc.)"),
        ("Malware Scanning", "Real-time malware scanning of downloaded content"),
        ("Logging", "Log all requests with user, URL, timestamp, action (allow/block)"),
        ("Bypass Prevention", "Block direct internet access, enforce proxy via firewall rules"),
        ("Redundancy", "High availability configuration (multiple proxy servers)"),
    ]
    
    row += 1
    for control, guidance in proxy_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:M{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["M"].width = 40


# ============================================================================
# SECTION 10: SHEET 7 - ADDITIONAL SERVICES
# ============================================================================

def create_additional_services_sheet(ws, styles, validations):
    """Create Additional Services assessment (Load Balancer, AAA, SNMP, Syslog, etc.)."""
    
    ws.title = "Additional_Services"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "Additional Network Services Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "📋 Assess other network services: Load Balancers, AAA (RADIUS/TACACS+), SNMP, Syslog, SMTP, FTP, etc."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "Service Type",         # B
        "Authentication",       # C - Y/N/NA
        "Encryption",           # D - Y/N/NA
        "Access Control",       # E - Y/N/NA
        "Logging",              # F - Y/N/NA
        "Monitoring",           # G - Y/N/NA
        "Hardening Applied",    # H - Y/N/NA
        "Compliance Score %",   # I - Auto
        "Last Assessed",        # J
        "Key Controls",         # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + ADDITIONAL_SERVICE_ROWS - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Service ID, Service Type
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Security controls
        for col in range(3, 9):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Compliance Score
        yes_range = f"C{row_idx}:H{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=9, value=formula)
        ws.cell(row=row_idx, column=9).number_format = "0.0"
        
        # Last Assessed, Key Controls, Notes
        for col in [10, 11, 12]:
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply validations
    for col in range(3, 9):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])
    
    for col in range(3, 9):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
    
    # Service-Specific Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "Service-Specific Security Controls Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Service Type"
    ws["B" + str(row)] = "Key Security Controls"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    service_controls = [
        ("Load Balancer", "SSL/TLS termination, strong ciphers, health checks, session persistence security, DDoS protection"),
        ("RADIUS/TACACS+", "Strong authentication, command authorization (TACACS+), encrypted communication, MFA support, logging"),
        ("SNMP", "SNMPv3 only (v1/v2c disabled), authentication (authPriv), encryption, access control, trap security"),
        ("Syslog", "TLS encryption (rsyslog/syslog-ng), log retention policy, log rotation, integrity protection"),
        ("SMTP Relay", "Authentication required, relay restrictions, SPF/DKIM/DMARC, TLS encryption, rate limiting"),
        ("FTP/SFTP", "Use SFTP (SSH), disable FTP, strong authentication, access control, logging"),
    ]
    
    row += 1
    for service, controls in service_controls:
        ws[f"A{row}"] = service
        ws.merge_cells(f"B{row}:L{row}")
        ws[f"B{row}"] = controls
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["K"].width = 35
    ws.column_dimensions["L"].width = 40


# ============================================================================
# SECTION 11: SHEET 8 - SERVICE COMPLIANCE SUMMARY
# ============================================================================

def create_service_compliance_summary_sheet(ws, styles):
    """Create Service Compliance Summary with metrics and charts."""
    
    ws.title = "Service_Compliance_Summary"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Service Security Compliance Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Overall Statistics
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Overall Service Statistics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    stats = [
        ("Total Services Cataloged", f'=COUNTA(Services_Catalog!B4:B{3 + SERVICE_ROW_COUNT})'),
        ("Critical Services", f'=COUNTIF(Services_Catalog!H4:H{3 + SERVICE_ROW_COUNT},"Critical")'),
        ("Services with Redundancy", f'=COUNTIF(Services_Catalog!I4:I{3 + SERVICE_ROW_COUNT},"Active-Active")+COUNTIF(Services_Catalog!I4:I{3 + SERVICE_ROW_COUNT},"Active-Passive")+COUNTIF(Services_Catalog!I4:I{3 + SERVICE_ROW_COUNT},"Clustered")'),
        ("Single Points of Failure", f'=COUNTIF(Services_Catalog!I4:I{3 + SERVICE_ROW_COUNT},"Single Point of Failure")'),
        ("Monitored Services", f'=COUNTIF(Services_Catalog!J4:J{3 + SERVICE_ROW_COUNT},"Monitored")'),
        ("Services Needing Review", f'=COUNTIF(Services_Catalog!M4:M{3 + SERVICE_ROW_COUNT},"<"&TODAY()-90)'),
    ]
    
    row += 1
    for metric, formula in stats:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Service Type Distribution
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Service Distribution by Type"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Service Type"
    ws["B" + str(row)] = "Count"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    chart_start_row = row + 1
    row += 1
    
    for service_type in SERVICE_TYPES:
        ws[f"A{row}"] = service_type
        ws[f"B{row}"] = f'=COUNTIF(Services_Catalog!B4:B{3 + SERVICE_ROW_COUNT},"{service_type}")'
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    chart_end_row = row - 1
    
    # Create Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Service Distribution by Type"
    pie_chart.style = 10
    labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_end_row)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.height = 10
    pie_chart.width = 15
    ws.add_chart(pie_chart, f"D{chart_start_row}")
    
    # Service Compliance Scores
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Average Compliance Scores by Service Type"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Service Type"
    ws["B" + str(row)] = "Avg Compliance %"
    ws["C" + str(row)] = "Services Assessed"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    row += 1
    compliance_data = [
        ("DNS", f'=AVERAGE(DNS_Security_Assessment!L4:L{3 + DNS_ASSESSMENT_ROWS})', f'=COUNTA(DNS_Security_Assessment!L4:L{3 + DNS_ASSESSMENT_ROWS})'),
        ("DHCP", f'=AVERAGE(DHCP_Security_Assessment!J4:J{3 + DHCP_ASSESSMENT_ROWS})', f'=COUNTA(DHCP_Security_Assessment!J4:J{3 + DHCP_ASSESSMENT_ROWS})'),
        ("NTP", f'=AVERAGE(NTP_Security_Assessment!H4:H{3 + NTP_ASSESSMENT_ROWS})', f'=COUNTA(NTP_Security_Assessment!H4:H{3 + NTP_ASSESSMENT_ROWS})'),
        ("Proxy", f'=AVERAGE(Proxy_Security_Assessment!J4:J{3 + PROXY_ASSESSMENT_ROWS})', f'=COUNTA(Proxy_Security_Assessment!J4:J{3 + PROXY_ASSESSMENT_ROWS})'),
        ("Other Services", f'=AVERAGE(Additional_Services!I4:I{3 + ADDITIONAL_SERVICE_ROWS})', f'=COUNTA(Additional_Services!I4:I{3 + ADDITIONAL_SERVICE_ROWS})'),
    ]
    
    for service_type, avg_formula, count_formula in compliance_data:
        ws[f"A{row}"] = service_type
        ws[f"B{row}"] = avg_formula
        ws[f"C{row}"] = count_formula
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws[f"B{row}"].number_format = "0.0"
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 3


# ============================================================================
# SECTION 12: SHEET 9 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(ws, styles, validations):
    """Create Gap Analysis sheet for service security gaps."""
    
    ws.title = "Gap_Analysis"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "Service Security Gaps - Remediation Tracking"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = '📋 Document all service security gaps (missing controls, misconfigurations). Track remediation progress.'
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",               # A - Auto
        "Service ID",           # B
        "Service Type",         # C
        "Security Gap",         # D
        "Current State",        # E
        "Severity",             # F - Dropdown
        "Remediation Plan",     # G
        "Owner",                # H
        "Status",               # I - Dropdown
        "Target Date",          # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Gap ID (auto-generated)
        gap_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SVC-GAP-" & TEXT({gap_num},"000"),"")')
        
        # Input cells
        for col in range(2, 11):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["gap_severity"].add(f"F{start_row}:F{end_row}")
    ws.add_data_validation(validations["gap_severity"])
    
    validations["gap_status"].add(f"I{start_row}:I{end_row}")
    ws.add_data_validation(validations["gap_status"])
    
    # Conditional Formatting for Severity
    
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 12
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 13: SHEET 10 - SERVICE DEPENDENCIES
# ============================================================================

def create_service_dependencies_sheet(ws, styles):
    """Create Service Dependencies mapping sheet."""
    
    ws.title = "Service_Dependencies"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Service Dependencies Mapping"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:F2")
    ws["A2"] = "📋 Map service dependencies to understand impact of service failures."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "Service Name",         # B
        "Depends On",           # C (Service IDs)
        "Required By",          # D (Service IDs)
        "Impact of Failure",    # E
        "Notes",                # F
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows (30 rows)
    start_row = 4
    end_row = start_row + 29
    
    for row_idx in range(start_row, end_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Example Dependencies
    row = end_row + 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Common Service Dependencies - Examples"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Service"
    ws["B" + str(row)] = "Typical Dependencies"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    dependencies = [
        ("Active Directory", "Depends On: DNS, NTP, DHCP (optional)"),
        ("Email (Exchange/O365)", "Depends On: DNS, Active Directory, Network Connectivity"),
        ("Web Applications", "Depends On: DNS, Load Balancer, Database, Authentication (LDAP/AD)"),
        ("VPN Access", "Depends On: RADIUS/TACACS+, Active Directory, DNS"),
        ("DHCP", "Depends On: Network connectivity (usually no dependencies)"),
        ("DNS", "Depends On: Network connectivity, NTP (for DNSSEC)"),
    ]
    
    row += 1
    for service, deps in dependencies:
        ws[f"A{row}"] = service
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = deps
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 40


# ============================================================================
# SECTION 14: MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.20-21-22 - Network Services Catalog Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.21 (Security of Network Services)")
    logger.info("=" * 80)
    logger.info("\n🎯 Systems Engineering Approach: Service-by-Service Security")
    logger.info("📊 Comprehensive Coverage: DNS, DHCP, NTP, Proxy, LB, AAA, SNMP, Syslog")
    logger.info("🔒 Audit-Ready: Compliance scoring and gap tracking")
    logger.info("\n" + "─" * 80)
    
    # Create workbook
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheet_names = [
        "Instructions & Guide",
        "Services_Catalog",
        "DNS_Security_Assessment",
        "DHCP_Security_Assessment",
        "NTP_Security_Assessment",
        "Proxy_Security_Assessment",
        "Additional_Services",
        "Service_Compliance_Summary",
        "Gap_Analysis",
        "Service_Dependencies",
    ]
    
    for name in sheet_names:
        wb.create_sheet(title=name)
    
    logger.info(f"✅ Workbook created with {len(sheet_names)} sheets")
    
    # Setup styles and validations
    logger.info("\n[Phase 2] Setting up styles and data validations...")
    styles = setup_styles()
    validations = create_data_validations()
    logger.info("✅ Styles and validations configured")
    
    # Create all sheets
    logger.info("\n[Phase 3] Generating assessment sheets...")
    
    logger.info("  [1/10] Creating Instructions & Guide...")
    create_instructions_sheet(wb["Instructions & Guide"], styles)
    logger.info("  ✅ Instructions complete")
    
    logger.info("  [2/10] Creating Services_Catalog...")
    create_services_catalog_sheet(wb["Services_Catalog"], styles, validations)
    logger.info(f"  ✅ Services catalog complete ({SERVICE_ROW_COUNT} service entries)")
    
    logger.info("  [3/10] Creating DNS_Security_Assessment...")
    create_dns_security_sheet(wb["DNS_Security_Assessment"], styles, validations)
    logger.info(f"  ✅ DNS assessment complete ({DNS_ASSESSMENT_ROWS} DNS services)")
    
    logger.info("  [4/10] Creating DHCP_Security_Assessment...")
    create_dhcp_security_sheet(wb["DHCP_Security_Assessment"], styles, validations)
    logger.info(f"  ✅ DHCP assessment complete ({DHCP_ASSESSMENT_ROWS} DHCP services)")
    
    logger.info("  [5/10] Creating NTP_Security_Assessment...")
    create_ntp_security_sheet(wb["NTP_Security_Assessment"], styles, validations)
    logger.info(f"  ✅ NTP assessment complete ({NTP_ASSESSMENT_ROWS} NTP services)")
    
    logger.info("  [6/10] Creating Proxy_Security_Assessment...")
    create_proxy_security_sheet(wb["Proxy_Security_Assessment"], styles, validations)
    logger.info(f"  ✅ Proxy assessment complete ({PROXY_ASSESSMENT_ROWS} proxy services)")
    
    logger.info("  [7/10] Creating Additional_Services...")
    create_additional_services_sheet(wb["Additional_Services"], styles, validations)
    logger.info(f"  ✅ Additional services complete ({ADDITIONAL_SERVICE_ROWS} services)")
    
    logger.info("  [8/10] Creating Service_Compliance_Summary...")
    create_service_compliance_summary_sheet(wb["Service_Compliance_Summary"], styles)
    logger.info("  ✅ Compliance summary complete")
    
    logger.info("  [9/10] Creating Gap_Analysis...")
    create_gap_analysis_sheet(wb["Gap_Analysis"], styles, validations)
    logger.info(f"  ✅ Gap analysis complete ({GAP_ROW_COUNT} gap tracking rows)")
    
    logger.info("  [10/10] Creating Service_Dependencies...")
    create_service_dependencies_sheet(wb["Service_Dependencies"], styles)
    logger.info("  ✅ Service dependencies complete")
    
    # Save workbook
    logger.info("\n[Phase 4] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.20-21-22.S3_Network_Services_Catalog_{GENERATED_TIMESTAMP}.xlsx"
    
    try:
        wb.save(filename)
        logger.info(f"✅ SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"❌ ERROR saving workbook: {e}")
        return 1
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("📋 WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\n📊 Assessment Sheets:")
    logger.info(f"  • Services_Catalog ({SERVICE_ROW_COUNT} master service entries)")
    logger.info(f"  • DNS_Security_Assessment ({DNS_ASSESSMENT_ROWS} rows, 9 controls)")
    logger.info(f"  • DHCP_Security_Assessment ({DHCP_ASSESSMENT_ROWS} rows, 7 controls)")
    logger.info(f"  • NTP_Security_Assessment ({NTP_ASSESSMENT_ROWS} rows, 5 controls)")
    logger.info(f"  • Proxy_Security_Assessment ({PROXY_ASSESSMENT_ROWS} rows, 7 controls)")
    logger.info(f"  • Additional_Services ({ADDITIONAL_SERVICE_ROWS} rows, 6 controls)")
    logger.info("\n📈 Analysis & Reporting:")
    logger.info("  • Service_Compliance_Summary (metrics, pie chart)")
    logger.info(f"  • Gap_Analysis ({GAP_ROW_COUNT} gap tracking rows)")
    logger.info("  • Service_Dependencies (dependency mapping)")
    logger.info("\n" + "=" * 80)
    logger.info("🚀 NEXT STEPS:")
    logger.info("  1. Complete Services_Catalog (all network services)")
    logger.info("  2. Assess each service type in dedicated sheets")
    logger.info("  3. Document gaps in Gap_Analysis")
    logger.info("  4. Map dependencies in Service_Dependencies")
    logger.info("  5. Review Service_Compliance_Summary")
    logger.info("\n" + "=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
