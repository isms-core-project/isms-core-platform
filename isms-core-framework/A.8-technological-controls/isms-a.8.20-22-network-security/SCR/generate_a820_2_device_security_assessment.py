#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-22.S2 - Network Device Security Assessment Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 2 of 6: Network Device Hardening and Security Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific hardening baselines, security standards, and
assessment requirements.

Key customization areas:
1. Hardening baseline requirements (match your security standards)
2. Device-specific security controls (vendor and device type specific)
3. Compliance scoring criteria (adapt to your risk profile)
4. Authentication requirements (based on your AAA infrastructure)
5. Gap severity classification (aligned with your risk assessment)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
network device security configurations against hardening baselines, supporting
evidence-based validation of device security controls.

**Purpose:**
Enables comprehensive security assessment of network devices against hardening
baselines, supporting evidence-based validation of device security configurations
and identification of gaps in compliance with ISO 27001:2022 Control A.8.20.

**Assessment Scope:**
- Device hardening baseline compliance
- Authentication mechanisms (local, AAA, 802.1X)
- Access control configuration (privilege levels, command authorisation)
- Encryption configuration (management protocols, wireless encryption)
- Logging and monitoring configuration
- Configuration backup status
- Patch/update status
- Gap identification and remediation tracking

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and hardening standards
2. Device_Security_Summary - Overall device security posture dashboard
3. Router_Hardening - Router-specific security assessment
4. Switch_Hardening - Switch-specific security assessment
5. Firewall_Hardening - Firewall-specific security assessment
6. Wireless_Hardening - Wireless AP-specific security assessment
7. LoadBalancer_Hardening - Load balancer-specific security assessment
8. VPN_Hardening - VPN concentrator-specific security assessment
9. Authentication_Controls - Authentication and access control assessment
10. Encryption_Controls - Encryption configuration assessment
11. Logging_Monitoring - Logging and monitoring configuration assessment
12. Gap_Analysis - Security gaps and remediation tracking
13. Evidence_Register - Audit evidence tracking and documentation
14. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with security control dropdown lists
- Conditional formatting for compliance status (compliant/non-compliant/partial)
- Automated compliance scoring per device
- Protected formulas with unprotected input cells
- Gap severity classification (Critical/High/Medium/Low)
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with configuration management systems

**Integration:**
This assessment imports device inventory from WB1 (Infrastructure Inventory)
and feeds security posture data into WB5 (Controls Coverage) and the Network
Security Compliance Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_2_device_security_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_2_device_security_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_2_device_security_assessment.py --date 20250124
    
    # Generate with custom organisation name
    python3 generate_a820_2_device_security_assessment.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organisation name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_2_Device_Security_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize hardening baselines to match your security standards
    2. Import device list from WB1 (Infrastructure Inventory)
    3. Assess each device against applicable hardening baseline
    4. Document security control implementation status
    5. Validate configurations against vendor hardening guides (CIS, vendor docs)
    6. Identify gaps and classify by severity
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (config exports, scan results)
    9. Obtain stakeholder approvals
    10. Feed results into WB5 and Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    2 of 6 (Network Device Hardening and Security Controls)
Primary Control:      A.8.20 (Network Security)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalisation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-22 specification
    - Supports comprehensive network device security assessment
    - Integrated with A.8.20-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Hardening Baseline Standards:**
Device hardening baselines should reference industry standards:
- CIS Benchmarks (Center for Internet Security)
- Vendor hardening guides (Cisco, Juniper, Palo Alto, etc.)
- DISA STIGs (for government/defense sectors)
- NIST SP 800-53 controls
- Organisation-specific security standards

Review and update baselines quarterly to incorporate new security requirements.

**Technology Diversity:**
This assessment framework is technology-agnostic and must work across:
- Traditional network devices (physical infrastructure)
- Virtual network appliances (NFV, virtual firewalls)
- Software-Defined Networks (SDN controllers, OpenFlow switches)
- Cloud-based network services (AWS, Azure, GCP)

Customize assessment criteria to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Documented hardening baselines for each device type
- Evidence of configuration compliance (config exports, scan results)
- Gap analysis with remediation plans
- Timelines and ownership for remediation activities
- Re-assessment results post-remediation

**Data Protection:**
Assessment workbooks contain sensitive security information including:
- Device security configurations and vulnerabilities
- Authentication and access control details
- Identified security gaps and weaknesses
- Remediation plans revealing security posture

Handle in accordance with your organisation's data classification policies.
Restrict access to authorised security personnel only.

**Maintenance:**
Review and update security assessments:
- Monthly: After device configuration changes or patches
- Quarterly: Routine reassessment of critical devices
- Annually: Complete reassessment of all devices
- Ad-hoc: After security incidents or vulnerability disclosures

**Quality Assurance:**
Validate assessment accuracy by:
- Exporting actual device configurations (show run, get config)
- Running automated compliance scanners (Nessus, Qualys, etc.)
- Manual verification of critical security controls
- Peer review by network security engineers
- Testing security controls (penetration testing)

**Regulatory Alignment:**
Ensure hardening standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 network device hardening requirements
- Healthcare: HIPAA network security configuration standards
- Finance: Regional banking network security requirements
- Government: Jurisdiction-specific hardening mandates

Customize assessment criteria to include regulatory-specific requirements.

**Integration Points:**
This assessment integrates with other ISO 27001 controls:
- A.8.8 (Vulnerability Management): Device vulnerability scanning results
- A.8.9 (Configuration Management): Configuration baseline compliance
- A.8.15 (Logging): Device logging configuration assessment
- A.8.16 (Monitoring): Device monitoring configuration assessment

**Common Pitfalls to Avoid:**
1. **Generic Baselines**: Using one-size-fits-all hardening (customize per device!)
2. **No Evidence**: Configuration claims without actual config exports
3. **Checkbox Compliance**: Marking "compliant" without verification
4. **Ignoring Legacy**: Excluding old devices from assessment
5. **No Remediation**: Identifying gaps without action plans
6. **Default Credentials**: Missing default password changes
7. **Unnecessary Services**: Not disabling unused protocols (Telnet, HTTP)
8. **Weak Encryption**: Allowing SSLv3, TLS 1.0, or weak ciphers

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "Network Device Security Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.20-22.S2"
CONTROL_ID   = "A.8.20-22"
CONTROL_NAME = "Network Security"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Assessment constants
DEVICE_ROW_COUNT = 51       # Device assessment rows (1 F2F2F2 sample + 50 FFFFCC)
GAP_ROW_COUNT = 100         # Gap tracking rows
REMEDIATION_ROW_COUNT = 50  # Remediation roadmap rows

DASH = "  -  "

# Excel formula sheet references (sheet names with spaces need single-quoting)
DHA_REF = "'Device Hardening Assessment'"  # Main assessment sheet
GS_REF = "'Gap Analysis'"                  # Gap summary sheet

# Hardening requirements (checklist items)
HARDENING_REQUIREMENTS = [
    "Default Credentials Disabled",
    "Strong Password Policy",
    "Multi-Factor Authentication (MFA)",
    "Unnecessary Services Disabled",
    "Secure Management (SSH/HTTPS Only)",
    "Logging Enabled",
    "NTP Configured",
    "SNMPv3 Only (v1/v2c Disabled)",
    "Session Timeouts Configured",
    "Banner Messages Configured",
    "Configuration Backups",
    "Firmware/Software Up-to-Date",
    "Access Control Lists (ACLs)",
    "Encrypted Management Traffic",
    "Least Privilege Access",
]

# Number of hardening requirements
NUM_REQUIREMENTS = len(HARDENING_REQUIREMENTS)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "yes_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="006100"),
        },
        "no_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="9C0006"),
        },
        "na_fill": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "font": Font(name="Calibri", size=10, color="7F7F7F"),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "low_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "noncompliant_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "partial_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "info_box": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Yes/No/N/A validation (for hardening requirements)
    validations["yes_no_na"] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False,
    )
    validations["yes_no_na"].error = "Must be Yes, No, or N/A"
    validations["yes_no_na"].errorTitle = "Invalid Value"
    
    # Device Type validation
    validations["device_type"] = DataValidation(
        type="list",
        formula1='"Router,Switch,Firewall,Wireless AP,Load Balancer,VPN Concentrator,IDS/IPS,Network Management,Other"',
        allow_blank=False,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False,
    )
    
    # Remediation Status validation
    validations["remediation_status"] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Completed,Accepted Risk,Deferred"',
        allow_blank=False,
    )
    
    # Compliance Status validation
    validations["compliance_status"] = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant,Partially Compliant,Not Assessed"',
        allow_blank=False,
    )
    
    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Device Hardening Assessment for each network device type in scope.', '2. Reference the Hardening Baseline Reference for vendor-specific standards (CIS, STIG, NIST).', '3. Use dropdown menus for standardised compliance and risk assessments.', '4. Document gaps in the Gap Analysis sheet with severity and remediation priority.', '5. Review the Compliance Scoring sheet for automated compliance percentage calculations.', '6. Assess device type compliance trends in the Device Type Compliance sheet.', '7. Prioritise top gaps in the Top Gaps Analysis sheet for remediation planning.', '8. Create remediation roadmaps with timelines in the Remediation Roadmap sheet.', '9. Maintain the Evidence Register with all supporting documentation for audit traceability.', '10. Obtain final approval and sign-off from IT Security, ISO, and CISO.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 24

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_device_hardening_assessment_sheet(ws, styles, validations):
    """Create main Device Hardening Assessment sheet."""
    
    ws.title = "Device Hardening Assessment"
    
    # Title
    title_col_span = 5 + NUM_REQUIREMENTS + 3  # Device info + requirements + scoring
    ws.merge_cells(f"A1:{get_column_letter(title_col_span)}1")
    cell = ws["A1"]
    cell.value = f"NETWORK DEVICE HARDENING ASSESSMENT - GENERATED {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells(f"A2:{get_column_letter(title_col_span)}2")
    ws["A2"] = f"Assess each device against {NUM_REQUIREMENTS} hardening requirements. Yellow cells are assessment fields (Yes/No/N/A). Compliance score auto-calculates."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Device ID",           # A (from WB1)
        "Device Type",         # B
        "Hostname",            # C
        "Primary IP",          # D
        "Criticality",         # E
    ]
    
    # Add hardening requirement columns
    for req in HARDENING_REQUIREMENTS:
        headers.append(req)
    
    # Add scoring columns
    headers.extend([
        "Compliance Score (%)",  # Auto-calculated
        "Gap Count",             # Auto-calculated
        "Assessment Date",
        "Assessor",
        "Notes",
    ])
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx <= 5 else 12
    
    # Data rows
    start_row = 4
    end_row = start_row + DEVICE_ROW_COUNT - 1
    
    req_start_col = 6  # Requirements start at column F
    req_end_col = 5 + NUM_REQUIREMENTS
    score_col = req_end_col + 1
    gap_col = score_col + 1
    
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _thin_s = Side(style="thin")
    _border_s = Border(left=_thin_s, right=_thin_s, top=_thin_s, bottom=_thin_s)

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        # Device info columns (A-E) - input cells
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _border_s
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

        # Hardening requirement columns (F onwards) - Yes/No/N/A dropdowns
        for col in range(req_start_col, req_end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _border_s
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

        # Compliance Score formula
        # Count "Yes" / (Count "Yes" + Count "No") * 100
        yes_range = f"{get_column_letter(req_start_col)}{row_idx}:{get_column_letter(req_end_col)}{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=score_col, value=formula)
        ws.cell(row=row_idx, column=score_col).number_format = "0.0"
        if is_sample:
            ws.cell(row=row_idx, column=score_col).fill = _sample_fill
            ws.cell(row=row_idx, column=score_col).border = _border_s

        # Gap Count formula (count "No" responses)
        gap_formula = f'=COUNTIF({yes_range},"No")'
        ws.cell(row=row_idx, column=gap_col, value=gap_formula)
        if is_sample:
            ws.cell(row=row_idx, column=gap_col).fill = _sample_fill
            ws.cell(row=row_idx, column=gap_col).border = _border_s

        # Assessment Date, Assessor, Notes
        for col in range(gap_col + 1, gap_col + 4):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _border_s
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Add sample data to row 4 for guidance
    ws.cell(row=start_row, column=1).value = "NET-DEV-001"
    ws.cell(row=start_row, column=2).value = "Router"
    ws.cell(row=start_row, column=3).value = "core-rtr-01"
    ws.cell(row=start_row, column=4).value = "192.168.1.1"
    ws.cell(row=start_row, column=5).value = "Critical"

    # Hardening requirement sample values (columns F-T = cols 6-20)
    req_sample_values = [
        "Yes",  # Default Credentials Disabled
        "Yes",  # Strong Password Policy
        "No",   # Multi-Factor Authentication (MFA) — gap for realism
        "Yes",  # Unnecessary Services Disabled
        "Yes",  # Secure Management (SSH/HTTPS Only)
        "Yes",  # Logging Enabled
        "Yes",  # NTP Configured
        "Yes",  # SNMPv3 Only (v1/v2c Disabled)
        "Yes",  # Session Timeouts Configured
        "Yes",  # Banner Messages Configured
        "Yes",  # Configuration Backups
        "No",   # Firmware/Software Up-to-Date — gap for realism
        "Yes",  # Access Control Lists (ACLs)
        "Yes",  # Encrypted Management Traffic
        "Yes",  # Least Privilege Access
    ]
    for col_offset, val in enumerate(req_sample_values):
        ws.cell(row=start_row, column=req_start_col + col_offset).value = val

    # Assessment date for sample row
    ws.cell(row=start_row, column=gap_col + 1).value = "15.01.2026"

    # Apply Data Validations to hardening requirement columns
    for col in range(req_start_col, req_end_col + 1):
        validations["yes_no_na"].add(
            f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}"
        )
    ws.add_data_validation(validations["yes_no_na"])
    
    # Conditional Formatting for Yes/No/N/A columns
    from openpyxl.formatting.rule import CellIsRule
    
    for col in range(req_start_col, req_end_col + 1):
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        
        # Yes = Green
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        # No = Red
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
        # N/A = Gray
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"N/A"'], fill=styles["na_fill"]["fill"], font=styles["na_fill"]["font"])
        )
    
    # Conditional Formatting for Compliance Score
    score_col_letter = get_column_letter(score_col)
    score_range = f"{score_col_letter}{start_row}:{score_col_letter}{end_row}"
    
    # >= 95% = Green (Compliant)
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["95"], fill=styles["compliant_fill"]["fill"])
    )
    # 80-94% = Yellow (Partially Compliant)
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="between", formula=["80", "94"], fill=styles["partial_fill"]["fill"])
    )
    # < 80% = Red (Non-Compliant)
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="lessThan", formula=["80"], fill=styles["noncompliant_fill"]["fill"])
    )

    # Freeze panes
    ws.freeze_panes = "A4"

    # Column widths for scoring columns
    ws.column_dimensions[get_column_letter(score_col)].width = 15
    ws.column_dimensions[get_column_letter(gap_col)].width = 10
    ws.column_dimensions[get_column_letter(gap_col + 1)].width = 15
    ws.column_dimensions[get_column_letter(gap_col + 2)].width = 20
    ws.column_dimensions[get_column_letter(gap_col + 3)].width = 40


# ============================================================================
# SECTION 6: SHEET 3 - HARDENING BASELINE REFERENCE
# ============================================================================

def create_hardening_baseline_reference_sheet(ws, styles):
    """Create Hardening Baseline Reference with CIS Benchmarks."""
    
    ws.title = "Hardening Baseline Reference"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "DEVICE HARDENING BASELINE REFERENCE"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Purpose
    ws.merge_cells("A2:F2")
    ws["A2"] = "Reference guide for device hardening based on CIS Benchmarks and vendor security guides"
    apply_style(ws["A2"], styles["info_box"])
    
    # CIS Benchmarks
    row = 4
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "CIS Benchmarks & Industry Standards"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 4):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Device Type"
    ws["B" + str(row)] = "Benchmark/Guide"
    ws["C" + str(row)] = "Version"
    ws["D" + str(row)] = "URL/Reference"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    benchmarks = [
        ("Cisco IOS", "CIS Cisco IOS Benchmark", "v4.1.0", "https://www.cisecurity.org/benchmark/cisco"),
        ("Juniper JunOS", "CIS Juniper Benchmark", "v2.1.0", "https://www.cisecurity.org/benchmark/juniper"),
        ("Palo Alto", "CIS Palo Alto Firewall Benchmark", "v1.1.0", "https://www.cisecurity.org/benchmark/palo_alto_networks"),
        ("Fortinet", "CIS Fortinet FortiOS Benchmark", "v1.1.0", "https://www.cisecurity.org/benchmark/fortinet"),
        ("Aruba", "Aruba Hardening Guide", "Latest", "https://www.arubanetworks.com/techdocs/"),
        ("Generic Network", "NIST SP 800-41 (Firewall/Router)", "Rev 1", "https://csrc.nist.gov/publications/detail/sp/800-41/rev-1/final"),
        ("Wireless", "CIS Wireless Network Security", "v1.0", "https://www.cisecurity.org/"),
    ]
    
    row += 1
    for device, benchmark, version, url in benchmarks:
        ws[f"A{row}"] = device
        ws[f"B{row}"] = benchmark
        ws[f"C{row}"] = version
        ws[f"D{row}"] = url
        for col in ["A", "B", "C", "D"]:
            apply_style(ws[f"{col}{row}"], styles["info_box"])
        row += 1
    
    # Hardening Requirements Detail
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Hardening Requirements - Implementation Guidance"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Requirement"
    ws["B" + str(row)] = "Implementation Guidance"
    ws["C" + str(row)] = "Verification Method"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    guidance = [
        (
            "Default Credentials Disabled",
            "Change default usernames/passwords (admin/admin, cisco/cisco, etc.). Disable or rename default accounts.",
            "Attempt login with default credentials (should fail). Review user list."
        ),
        (
            "Strong Password Policy",
            "Min 12 chars, complexity (upper/lower/number/special), expiration 90 days, history 5 passwords.",
            "Check password policy config. Test weak password (should be rejected)."
        ),
        (
            "Multi-Factor Authentication",
            "Enable MFA for admin access (TACACS+, RADIUS with OTP/tokens). Critical for privileged access.",
            "Verify AAA config, test login (should require 2nd factor)."
        ),
        (
            "Unnecessary Services Disabled",
            "Disable unused protocols: HTTP (use HTTPS), Telnet (use SSH), FTP, TFTP, finger, echo, CDP (on edge).",
            "Port scan device, check running services list."
        ),
        (
            "Secure Management (SSH/HTTPS)",
            "Use SSH for CLI (disable Telnet), HTTPS for web (disable HTTP). SSH v2 only.",
            "Telnet to device (should fail). Check SSH version."
        ),
        (
            "Logging Enabled",
            "Enable syslog to centralised server. Log severity: informational or higher. Buffer logging.",
            "Check logging config, verify logs reaching syslog server."
        ),
        (
            "NTP Configured",
            "Configure NTP client, point to internal NTP servers. Enable NTP authentication if possible.",
            "Check NTP config, verify time sync (ntpstat, show ntp status)."
        ),
        (
            "SNMPv3 Only",
            "Use SNMPv3 with authentication and encryption. Disable SNMPv1/v2c (community strings).",
            "Check SNMP config, test v2c query (should fail)."
        ),
        (
            "Session Timeouts",
            "Configure idle timeout (e.g., 10 minutes for console, 15 min for VTY). Auto-logout inactive sessions.",
            "Check timeout config (exec-timeout, idle-timeout)."
        ),
        (
            "Banner Messages",
            "Configure login banner warning unauthorised access is prohibited. MOTD and login banners.",
            "Check banner config, verify displayed on login."
        ),
        (
            "Configuration Backups",
            "Automated nightly config backups to secure location. Use RANCID, Oxidized, or scripts.",
            "Verify backup schedule, check recent backup files exist."
        ),
        (
            "Firmware/Software Updated",
            "Running latest stable firmware/software with security patches. Track end-of-life dates.",
            "Check current version vs. vendor latest. Review security advisories."
        ),
        (
            "Access Control Lists",
            "Implement ACLs on routers/firewalls for ingress/egress filtering. Default deny, explicit allow.",
            "Review ACL config, test traffic (permitted/denied as expected)."
        ),
        (
            "Encrypted Management",
            "Encrypt management traffic: SSH (not Telnet), HTTPS (not HTTP), SNMPv3, TLS syslog.",
            "Capture management traffic (tcpdump), verify encrypted."
        ),
        (
            "Least Privilege Access",
            "Role-based access control (RBAC). Users have minimum privileges needed. Privilege levels configured.",
            "Review user privilege levels, test restricted user access."
        ),
    ]
    
    row += 1
    for requirement, impl, verify in guidance:
        ws[f"A{row}"] = requirement
        ws[f"B{row}"] = impl
        ws[f"C{row}"] = verify
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        row += 1
    
    # Device-Specific Notes
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Device-Specific Hardening Notes"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    device_notes = [
        ("Routers", "Disable IP source routing, enable uRPF, implement ACLs, secure routing protocols (MD5 auth)."),
        ("Switches", "Enable port security, DHCP snooping, DAI, disable DTP, change native VLAN, disable unused ports."),
        ("Firewalls", "Default deny policy, rule cleanup, logging all denied traffic, regular rule review."),
        ("Wireless APs", "WPA3 or WPA2-Enterprise, disable WPS, hide SSID (optional), MAC filtering, rogue AP detection."),
        ("Load Balancers", "Strong SSL/TLS ciphers, disable SSLv3/TLS1.0, enable HSTS, session persistence security."),
    ]
    
    for device, notes in device_notes:
        ws[f"A{row}"] = device
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = notes
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


# ============================================================================
# SECTION 7: SHEET 4 - GAP SUMMARY
# ============================================================================

def create_gap_summary_sheet(ws, styles, validations):
    """Create Gap Analysis sheet for identified hardening gaps."""
    
    ws.title = "Gap Analysis"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "DEVICE HARDENING GAPS - SUMMARY & REMEDIATION TRACKING"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = 'Document all "No" responses from Device Hardening Assessment. Track remediation progress and priority.'
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",                # A - Auto
        "Device ID",             # B
        "Device Type",           # C
        "Hostname",              # D
        "Hardening Requirement", # E (which requirement failed)
        "Current State",         # F (description)
        "Gap Severity",          # G (Critical/High/Medium/Low)
        "Remediation Plan",      # H
        "Owner",                 # I
        "Status",                # J (dropdown)
        "Target Date",           # K
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1

    _sample_fill_gs = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _thin_gs = Side(style="thin")
    _border_gs = Border(left=_thin_gs, right=_thin_gs, top=_thin_gs, bottom=_thin_gs)

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        # Gap ID (auto-generated)
        gap_num = row_idx - start_row + 1
        cell_a = ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","HARD-GAP-" & TEXT({gap_num},"000"),"")')
        if is_sample:
            cell_a.fill = _sample_fill_gs
            cell_a.border = _border_gs
            cell_a.font = Font(color="808080", name="Calibri")
        else:
            cell_a.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell_a.border = _border_gs

        # Input cells
        for col in range(2, 12):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill_gs
                cell.border = _border_gs
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample row data
    ws.cell(row=start_row, column=1).value = "HARD-GAP-001"  # override auto-formula
    ws.cell(row=start_row, column=2).value = "NET-DEV-001"
    ws.cell(row=start_row, column=3).value = "Router"
    ws.cell(row=start_row, column=4).value = "core-rtr-01"
    ws.cell(row=start_row, column=5).value = "Multi-Factor Authentication (MFA)"
    ws.cell(row=start_row, column=6).value = "MFA not configured for admin access"
    ws.cell(row=start_row, column=7).value = "Critical"
    ws.cell(row=start_row, column=8).value = "Configure TACACS+ with MFA within 30 days"
    ws.cell(row=start_row, column=9).value = "Network Security Team"
    ws.cell(row=start_row, column=10).value = "Open"
    ws.cell(row=start_row, column=11).value = "28.02.2026"

    # Apply Data Validations (exclude sample row 4 from DV — start from row 5)
    validations["gap_severity"].add(f"G{start_row + 1}:G{end_row}")
    ws.add_data_validation(validations["gap_severity"])

    validations["remediation_status"].add(f"J{start_row + 1}:J{end_row}")
    ws.add_data_validation(validations["remediation_status"])
    
    # Conditional Formatting for Severity
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Gap Severity Guidelines
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "Gap Severity Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 12):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Severity"
    ws["B" + str(row)] = "Definition"
    ws["C" + str(row)] = "Examples"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    severity_guide = [
        ("Critical", "Immediate exploitation risk, default credentials, no encryption", "Default passwords still enabled, Telnet enabled on Internet-facing device"),
        ("High", "Significant risk, missing key controls", "No MFA for admin access, SNMPv1/v2c enabled, no logging"),
        ("Medium", "Moderate risk, important controls missing", "Weak password policy, no session timeouts, outdated firmware (non-critical)"),
        ("Low", "Minor gap, limited risk", "Missing login banner, unused ports not disabled"),
    ]
    
    row += 1
    for severity, defn, examples in severity_guide:
        ws[f"A{row}"] = severity
        ws[f"B{row}"] = defn
        ws[f"C{row}"] = examples
        
        if severity == "Critical":
            apply_style(ws[f"A{row}"], styles["critical_fill"])
        elif severity == "High":
            apply_style(ws[f"A{row}"], styles["high_fill"])
        elif severity == "Medium":
            apply_style(ws[f"A{row}"], styles["medium_fill"])
        elif severity == "Low":
            apply_style(ws[f"A{row}"], styles["low_fill"])
        
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 12
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SHEET 5 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1, TABLE 2, TABLE 3."""

    ws.title = "Summary Dashboard"

    SQ = "'"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _white_bold_14 = Font(bold=True, color="FFFFFF", size=14)
    _white_bold_11 = Font(bold=True, color="FFFFFF", size=11)
    _dark_bold_10 = Font(bold=True, color="000000", size=10)
    _dark_10 = Font(bold=False, color="000000", size=10)
    _dark_9_italic = Font(bold=False, color="000000", size=9, italic=True)
    _title_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _data_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _t3total_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _t3_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    _center = Alignment(horizontal="center", vertical="center")
    _left = Alignment(horizontal="left", vertical="center")
    _wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _banner(row, text, font, fill):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = font
        c.fill = fill

    def _hdr_row_t1(row):
        for col_letter, text in [("A", "Assessment Area"), ("B", "Total"), ("C", "Yes"), ("D", "Partial"), ("E", "No"), ("F", "N-A"), ("G", "Compliance %")]:
            c = ws[f"{col_letter}{row}"]
            c.value = text
            c.font = _dark_bold_10
            c.fill = _hdr_fill
            c.border = _thin
            c.alignment = _center

    def _data_row(row, area, b, c_val, d, e, f_val, g):
        ws[f"A{row}"].value = area
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col, val in [("B", b), ("C", c_val), ("D", d), ("E", e), ("F", f_val)]:
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.font = _dark_10
            cell.border = _thin
            cell.alignment = _center
        ws[f"G{row}"].value = g
        ws[f"G{row}"].font = _dark_10
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    def _total_row_t1(row, first, last):
        ws[f"A{row}"].value = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"A{row}"].fill = _total_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col in ["B", "C", "D", "E", "F"]:
            c = ws[f"{col}{row}"]
            c.value = f"=SUM({col}{first}:{col}{last})"
            c.font = Font(bold=True, color="000000", size=10)
            c.fill = _total_fill
            c.border = _thin
            c.alignment = _center
        ws[f"G{row}"].value = f"=IF((B{row}-F{row})=0,\"0%\",ROUND(C{row}/(B{row}-F{row})*100,1)&\"%\")"
        ws[f"G{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"G{row}"].fill = _total_fill
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    # ── ROW 1: Title ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 35
    _banner(1, "NETWORK DEVICE SECURITY ASSESSMENT — SUMMARY DASHBOARD", _white_bold_14, _title_fill)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── ROW 2: Subtitle ───────────────────────────────────────────────
    ws["A2"] = "ISO/IEC 27001:2022 — Controls A.8.20 · A.8.21 · A.8.22: Network Security, Security of Network Services and Segregation of Networks"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = _left
    ws.merge_cells("A2:G2")

    # ── TABLE 1 ───────────────────────────────────────────────────────
    _banner(4, "TABLE 1: ASSESSMENT AREA COMPLIANCE", _white_bold_11, _title_fill)
    _hdr_row_t1(5)

    # Row 6: Device Hardening Assessment (score col U, rows 4-54, DEVICE_ROW_COUNT=51)
    # Scoring: >=80 = Yes (compliant), 50-79 = Partial, <50 = No (non-compliant)
    _data_row(6,
        "Device Hardening Assessment",
        "=COUNTA('Device Hardening Assessment'!B5:B54)",
        "=COUNTIF('Device Hardening Assessment'!U5:U54,\">=80\")",
        "=COUNTIFS('Device Hardening Assessment'!U5:U54,\">=50\",'Device Hardening Assessment'!U5:U54,\"<80\")",
        "=COUNTIFS('Device Hardening Assessment'!U5:U54,\">=0\",'Device Hardening Assessment'!U5:U54,\"<50\")",
        "=B6-(C6+D6+E6)",
        "=IF((B6-F6)=0,\"0%\",ROUND(C6/(B6-F6)*100,1)&\"%\")"
    )

    # Row 7: Gap Analysis (status col J, rows 5-110)
    # DV values: Open / In Progress / Completed / Accepted Risk / Deferred
    _data_row(7,
        "Gap Analysis",
        "=COUNTA('Gap Analysis'!B5:B110)",
        "=COUNTIF('Gap Analysis'!J5:J110,\"Completed\")",
        "=COUNTIF('Gap Analysis'!J5:J110,\"In Progress\")",
        "=COUNTIF('Gap Analysis'!J5:J110,\"Open\")",
        "=B7-(C7+D7+E7)",
        "=IF((B7-F7)=0,\"0%\",ROUND(C7/(B7-F7)*100,1)&\"%\")"
    )

    # Row 8: TOTAL
    _total_row_t1(8, 6, 7)

    # ── TABLE 2 ───────────────────────────────────────────────────────
    _banner(10, "TABLE 2: KEY METRICS", _white_bold_11, _title_fill)

    ws["A11"].value = "Metric"
    ws["A11"].font = _dark_bold_10
    ws["A11"].fill = _hdr_fill
    ws["A11"].border = _thin
    ws["A11"].alignment = _left
    ws["B11"].value = "Value"
    ws["B11"].font = _dark_bold_10
    ws["B11"].fill = _hdr_fill
    ws["B11"].border = _thin
    ws["B11"].alignment = _center
    ws.merge_cells("C11:G11")
    ws["C11"].value = "What This Shows"
    ws["C11"].font = _dark_bold_10
    ws["C11"].fill = _hdr_fill
    ws["C11"].border = _thin
    ws["C11"].alignment = _left

    t2_metrics = [
        (12, "Non-Compliant Devices (<80%)",
         "=COUNTIFS('Device Hardening Assessment'!U5:U54,\">=0\",'Device Hardening Assessment'!U5:U54,\"<80\")",
         "Devices scoring below 80% hardening baseline (require remediation)"),
        (13, "Fully Compliant Devices (\u226595%)",
         "=COUNTIF('Device Hardening Assessment'!U5:U54,\">=95\")",
         "Devices meeting gold-standard hardening (\u226595% compliance score)"),
        (14, "Partially Compliant Devices (80\u201394%)",
         "=COUNTIFS('Device Hardening Assessment'!U5:U54,\">=80\",'Device Hardening Assessment'!U5:U54,\"<95\")",
         "Near-compliant devices needing minor hardening improvements"),
        (15, "Critical Severity Hardening Gaps",
         "=COUNTIF('Gap Analysis'!G5:G110,\"Critical\")",
         "Critical hardening gaps requiring immediate remediation"),
        (16, "Open Gaps",
         "=COUNTIF('Gap Analysis'!J5:J110,\"Open\")",
         "Outstanding gaps without active remediation plan"),
        (17, "High Severity Gaps",
         "=COUNTIF('Gap Analysis'!G5:G110,\"High\")",
         "High-severity hardening gaps requiring prioritised remediation"),
        (18, "Devices Not Assessed",
         "=COUNTBLANK('Device Hardening Assessment'!U5:U54)",
         "Devices in scope with no hardening score (assessment required)"),
        (19, "Overall Compliance Rate",
         "=G8",
         "Overall device hardening compliance percentage (TABLE 1 total)"),
    ]

    for row, metric, formula, description in t2_metrics:
        ws[f"A{row}"].value = metric
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = description
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).border = _thin

    # ── TABLE 3 ───────────────────────────────────────────────────────
    _banner(21, "TABLE 3: CRITICAL FINDINGS", _white_bold_11, _t3_fill)

    ws["A22"].value = "Critical Finding Type"
    ws["A22"].font = _dark_bold_10
    ws["A22"].fill = _hdr_fill
    ws["A22"].border = _thin
    ws["A22"].alignment = _left
    ws["B22"].value = "Count"
    ws["B22"].font = _dark_bold_10
    ws["B22"].fill = _hdr_fill
    ws["B22"].border = _thin
    ws["B22"].alignment = _center
    ws.merge_cells("C22:G22")
    ws["C22"].value = "Filter Instructions"
    ws["C22"].font = _dark_bold_10
    ws["C22"].fill = _hdr_fill
    ws["C22"].border = _thin
    ws["C22"].alignment = _left

    t3_findings = [
        (23, "Non-Compliant Devices (<80%)",
         "=COUNTIFS('Device Hardening Assessment'!U5:U54,\">=0\",'Device Hardening Assessment'!U5:U54,\"<80\")",
         "Filter Device Hardening Assessment: Score < 80"),
        (24, "Critical Hardening Gaps",
         "=COUNTIF('Gap Analysis'!G5:G110,\"Critical\")",
         "Filter Gap Analysis: Severity = \"Critical\""),
        (25, "Open Gaps",
         "=COUNTIF('Gap Analysis'!J5:J110,\"Open\")",
         "Filter Gap Analysis: Status = \"Open\""),
        (26, "High Severity Gaps",
         "=COUNTIF('Gap Analysis'!G5:G110,\"High\")",
         "Filter Gap Analysis: Severity = \"High\""),
        (27, "Devices Not Assessed",
         "=COUNTBLANK('Device Hardening Assessment'!U5:U54)",
         "Filter Device Hardening Assessment: Score is empty"),
    ]

    for row, finding_type, formula, instructions in t3_findings:
        ws[f"A{row}"].value = finding_type
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].fill = _data_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].fill = _data_fill
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = instructions
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].fill = _data_fill
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).fill = _data_fill
            ws.cell(row=row, column=_c).border = _thin

    # TOTAL row
    ws["A28"].value = "TOTAL"
    ws["A28"].font = Font(bold=True, color="000000", size=10)
    ws["A28"].border = _thin
    ws["A28"].alignment = _left
    ws["B28"].value = "=SUM(B23:B27)"
    ws["B28"].font = Font(bold=True, color="000000", size=10)
    ws["B28"].fill = _t3total_fill
    ws["B28"].border = _thin
    ws["B28"].alignment = _center
    ws.merge_cells("C28:G28")
    ws["C28"].value = "Total critical findings requiring immediate remediation"
    ws["C28"].font = Font(italic=True, size=9, color="000000")
    ws["C28"].alignment = _left

    # ── Column widths ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 9: SHEET 6 - DEVICE TYPE COMPLIANCE
# ============================================================================

def create_device_type_compliance_sheet(ws, styles):
    """Create Device Type Compliance breakdown."""
    
    ws.title = "Device Type Compliance"
    
    # Title
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "COMPLIANCE ANALYSIS BY DEVICE TYPE"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35

    # Purpose
    ws.merge_cells("A2:E2")
    ws["A2"] = "Breakdown of hardening compliance scores by device type to identify problem areas"
    apply_style(ws["A2"], styles["info_box"])

    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Average Compliance Score by Device Type"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 6):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Device Type"
    ws["B" + str(row)] = "Device Count"
    ws["C" + str(row)] = "Avg Compliance Score"
    ws["D" + str(row)] = "Min Score"
    ws["E" + str(row)] = "Max Score"
    for col in ["A", "B", "C", "D", "E"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    device_types = [
        "Router",
        "Switch",
        "Firewall",
        "Wireless AP",
        "Load Balancer",
        "VPN Concentrator",
        "IDS/IPS",
        "Network Management",
        "Other",
    ]
    
    # Get score column from main sheet
    req_end = 5 + NUM_REQUIREMENTS
    score_col = req_end + 1
    score_col_letter = get_column_letter(score_col)
    
    chart_start_row = row + 1
    row += 1
    
    for dev_type in device_types:
        ws[f"A{row}"] = dev_type
        ws[f"B{row}"] = f'=COUNTIF({DHA_REF}!B5:B{3 + DEVICE_ROW_COUNT},"{dev_type}")'
        ws[f"C{row}"] = f'=IFERROR(AVERAGEIF({DHA_REF}!B5:B{3 + DEVICE_ROW_COUNT},"{dev_type}",{DHA_REF}!{score_col_letter}5:{score_col_letter}{3 + DEVICE_ROW_COUNT}),"")'
        ws[f"D{row}"] = f'=IFERROR(MINIFS({DHA_REF}!{score_col_letter}5:{score_col_letter}{3 + DEVICE_ROW_COUNT},{DHA_REF}!B5:B{3 + DEVICE_ROW_COUNT},"{dev_type}"),"")'
        ws[f"E{row}"] = f'=IFERROR(MAXIFS({DHA_REF}!{score_col_letter}5:{score_col_letter}{3 + DEVICE_ROW_COUNT},{DHA_REF}!B5:B{3 + DEVICE_ROW_COUNT},"{dev_type}"),"")'
        
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        apply_style(ws[f"E{row}"], styles["info_box"])
        
        ws[f"C{row}"].number_format = "0.0"
        ws[f"D{row}"].number_format = "0.0"
        ws[f"E{row}"].number_format = "0.0"
        
        row += 1
    
    chart_end_row = row - 1
    
    # Create Bar Chart - Compliance by Device Type
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Average Compliance Score by Device Type"
    bar_chart.y_axis.title = "Compliance Score (%)"
    bar_chart.x_axis.title = "Device Type"
    labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    data = Reference(ws, min_col=3, min_row=chart_start_row, max_row=chart_end_row)
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.height = 12
    bar_chart.width = 18
    ws.add_chart(bar_chart, f"A{row + 2}")
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 3


# ============================================================================
# SECTION 10: SHEET 7 - TOP GAPS ANALYSIS
# ============================================================================

def create_top_gaps_analysis_sheet(ws, styles):
    """Create Top Gaps Analysis showing most common failures."""
    
    ws.title = "Top Gaps Analysis"
    
    # Title
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "TOP HARDENING GAPS - MOST COMMON FAILURES"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Purpose
    ws.merge_cells("A2:E2")
    ws["A2"] = "Identify most frequently failed hardening requirements across all devices"
    apply_style(ws["A2"], styles["info_box"])
    
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"Hardening Requirements Failure Analysis ({NUM_REQUIREMENTS} Requirements)"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 6):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Requirement"
    ws["B" + str(row)] = "Total 'No' Count"
    ws["C" + str(row)] = "Failure Rate %"
    ws["D" + str(row)] = "Priority"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    # Calculate failure counts for each requirement
    req_start_col = 6
    chart_start_row = row + 1
    row += 1
    
    for idx, req in enumerate(HARDENING_REQUIREMENTS):
        col_letter = get_column_letter(req_start_col + idx)
        ws[f"A{row}"] = req
        ws[f"B{row}"] = f'=COUNTIF({DHA_REF}!{col_letter}5:{col_letter}{3 + DEVICE_ROW_COUNT},"No")'
        ws[f"C{row}"] = f'=IFERROR(B{row}/COUNTA({DHA_REF}!{col_letter}5:{col_letter}{3 + DEVICE_ROW_COUNT}),"")'
        ws[f"D{row}"] = f'=IF(C{row}>=75,"CRITICAL",IF(C{row}>=50,"HIGH",IF(C{row}>=25,"MEDIUM","LOW")))'
        
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        
        ws[f"C{row}"].number_format = "0.0%"
        
        row += 1
    
    chart_end_row = row - 1
    
    # Create Bar Chart - Top Gaps
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Most Common Hardening Failures"
    bar_chart.y_axis.title = "Failure Count"
    bar_chart.x_axis.title = "Requirement"
    labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_end_row)
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.height = 15
    bar_chart.width = 20
    ws.add_chart(bar_chart, f"A{row + 2}")
    
    # Conditional Formatting for Priority
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"HIGH"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"D{chart_start_row}:D{chart_end_row}",
        CellIsRule(operator="equal", formula=['"LOW"'], fill=styles["low_fill"]["fill"])
    )
    
    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 3


# ============================================================================
# SECTION 11: SHEET 8 - REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap_sheet(ws, styles, validations):
    """Create Remediation Roadmap with prioritised action plan."""
    
    ws.title = "Remediation Roadmap"
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "HARDENING REMEDIATION ROADMAP"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "Prioritised remediation plan sorted by severity. Focus on Critical/High gaps first."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Priority",            # A (1, 2, 3...)
        "Severity",            # B (Critical/High/Medium/Low)
        "Gap Description",     # C
        "Affected Devices",    # D (count or list)
        "Remediation Action",  # E
        "Owner",               # F
        "Target Date",         # G
        "Status",              # H (dropdown)
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + REMEDIATION_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Priority (manual numbering)
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        
        # Input cells
        for col in range(2, 9):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["gap_severity"].add(f"B{start_row}:B{end_row}")
    ws.add_data_validation(validations["gap_severity"])
    
    validations["remediation_status"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["remediation_status"])
    
    # Conditional Formatting for Severity
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Remediation Approach
    row = end_row + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Remediation Approach Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 9):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    approach = [
        "1. PRIORITIZE BY SEVERITY: Address Critical gaps immediately, then High, then Medium/Low",
        "2. CLUSTER BY REQUIREMENT: Remediate same requirement across multiple devices together (efficiency)",
        "3. LOW-HANGING FRUIT: Quick wins (e.g., enable logging, configure banners) → build momentum",
        "4. COMPLEX ITEMS: MFA, major config changes require planning, testing, approval",
        "5. VALIDATE: Test in lab/dev environment before production deployment",
        "6. DOCUMENT: Maintain change records, configuration backups, validation evidence",
        "7. RE-ASSESS: After remediation, re-run assessment to confirm compliance improvement",
    ]
    
    for item in approach:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = item
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 12: EVIDENCE REGISTER (NEW - GOLDEN STANDARD)
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (italic, NOT blue banner) ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers (003366, white text) ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Device Security Assessment",
        3: "Configuration file",
        4: "Device hardening assessment results and security configuration exports",
        5: "/evidence/A.8.20-22/",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet -- golden standard common sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G7),\"\")"),
        ("Assessment Status:", ""),
    ]
    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    status_dv = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=True)
    ws.add_data_validation(status_dv)
    if status_row:
        status_dv.add(ws[f"B{status_row}"])

    approvers = [
        ("COMPLETED BY (NETWORK SECURITY)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True, name="Calibri")
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    decision_dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 14: MAIN FUNCTION
# ============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    styles = _STYLES
    validations = create_data_validations()

    create_instructions_sheet(wb.create_sheet())
    create_device_hardening_assessment_sheet(wb.create_sheet(), styles, validations)
    create_hardening_baseline_reference_sheet(wb.create_sheet(), styles)
    create_gap_summary_sheet(wb.create_sheet(), styles, validations)
    create_device_type_compliance_sheet(wb.create_sheet(), styles)
    create_top_gaps_analysis_sheet(wb.create_sheet(), styles)
    create_remediation_roadmap_sheet(wb.create_sheet(), styles, validations)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    create_summary_dashboard_sheet(wb.create_sheet(), styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_sheet(ws, styles)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path.name}")

def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"ERROR saving workbook: {e}")
        sys.exit(1)
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\nAssessment Sheets:")
    logger.info("  • Instructions & Legend (methodology, assessment criteria)")
    logger.info(f"  • Device Hardening Assessment ({DEVICE_ROW_COUNT} devices × {NUM_REQUIREMENTS} requirements)")
    logger.info("  • Hardening Baseline Reference (CIS Benchmarks, vendor guides)")
    logger.info(f"  • Gap Analysis ({GAP_ROW_COUNT} gap tracking rows)")
    logger.info("\n Analysis & Reporting:")
    logger.info("  • Compliance Scoring (overall metrics, targets, bar chart)")
    logger.info("  • Device Type Compliance (breakdown by device type, bar chart)")
    logger.info("  • Top Gaps Analysis (most common failures, bar chart)")
    logger.info(f"  • Remediation Roadmap ({REMEDIATION_ROW_COUNT} prioritised actions)")
    logger.info("\n" + "─" * 80)
    logger.info(" ASSESSMENT CAPABILITIES:")
    logger.info(f"  • {DEVICE_ROW_COUNT} device hardening assessments")
    logger.info(f"  • {NUM_REQUIREMENTS} hardening requirements per device")
    logger.info("  • Auto-calculated compliance scores (Yes/No/N/A → %)")
    logger.info("  • Conditional formatting (Yes=Green, No=Red, N/A=Gray)")
    logger.info(f"  • {GAP_ROW_COUNT} gap identification and tracking")
    logger.info(f"  • {REMEDIATION_ROW_COUNT} remediation roadmap items")
    logger.info("  • 3 charts (compliance pie, device type bar, top gaps bar)")
    logger.info("  • CIS Benchmark references and implementation guidance")
    logger.info("\n" + "─" * 80)
    logger.info(" KEY FEATURES:")
    logger.info("  Based on CIS Benchmarks and vendor hardening guides")
    logger.info("  Systematic Yes/No/N/A assessment per requirement")
    logger.info("  Automated compliance scoring formulas")
    logger.info("  Gap identification with severity classification")
    logger.info("  Remediation tracking and prioritization")
    logger.info("  Device type compliance analysis")
    logger.info("  Top gaps visualization")
    logger.info("  Integration with Workbook 1 (Device Inventory)")
    logger.info("\n" + "=" * 80)
    logger.info("NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Read Instructions & Legend sheet first")
    logger.info("  3. Review Hardening Baseline Reference (CIS Benchmarks)")
    logger.info("  4. For each device from WB1, assess hardening compliance")
    logger.info("  5. Fill Device Hardening Assessment (Yes/No/N/A per requirement)")
    logger.info("  6. Document all 'No' responses in Gap Analysis")
    logger.info("  7. Review Compliance Scoring and Device Type Compliance")
    logger.info("  8. Prioritize remediation using Remediation Roadmap")
    logger.info("  9. Re-assess after remediation to track improvement")
    logger.info("\nPRO TIP:")
    logger.info("  Start with Critical devices first (from WB1 criticality assessment).")
    logger.info("  Use Hardening Baseline Reference as your assessment guide.")
    logger.info("  Don't mark everything 'N/A' - be honest about gaps to improve security.")
    logger.info("\n" + "=" * 80)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info("\n This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
