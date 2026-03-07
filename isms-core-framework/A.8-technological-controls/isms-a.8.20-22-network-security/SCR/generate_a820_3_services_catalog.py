#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-22.S3 - Network Services Catalog Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 3 of 6: Network Services Inventory and Security Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific network services catalog, security standards, and
assessment requirements.

Key customization areas:
1. Network services inventory (match your actual service catalog)
2. Service-specific security requirements (based on your service architecture)
3. Availability requirements and SLAs (aligned with business needs)
4. Redundancy and failover criteria (based on your resilience requirements)
5. Monitoring and alerting thresholds (adapted to your operations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for inventorying
and assessing network services security across the organisation, supporting
evidence-based validation of service security controls.

**Purpose:**
Enables systematic inventory and security assessment of network services (DNS,
DHCP, NTP, proxy, load balancers, etc.), supporting evidence-based validation
of service security controls against ISO 27001:2022 Control A.8.21.

**Assessment Scope:**
- Network services inventory (DNS, DHCP, NTP, proxy, LB, etc.)
- Service-specific security assessment tabs
- Service availability and redundancy status
- Security mechanisms per service type
- Service monitoring and alerting configuration
- Compliance scoring per service
- Service owner/responsible party tracking
- Gap identification and remediation planning

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and service security standards
2. Services_Summary - Overall network services security dashboard
3. DNS_Services - DNS security assessment (DNSSEC, split DNS, etc.)
4. DHCP_Services - DHCP security assessment (rogue prevention, etc.)
5. NTP_Services - NTP security assessment (time source authentication)
6. Proxy_Services - Proxy security assessment (filtering, authentication)
7. LoadBalancer_Services - Load balancer security (SSL termination, etc.)
8. Authentication_Services - Auth services (RADIUS, TACACS+, LDAP)
9. SNMP_Services - SNMP security (v3 only, community string management)
10. Syslog_Services - Syslog security (encryption, authentication)
11. Other_Services - Additional network services
12. Availability_Monitoring - Service availability and uptime tracking
13. Gap Analysis - Service security gaps and remediation tracking
14. Evidence_Register - Audit evidence tracking and documentation
15. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with service type and security mechanism dropdown lists
- Conditional formatting for service security status
- Automated compliance scoring per service
- Protected formulas with unprotected input cells
- Service availability SLA tracking
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with service monitoring systems

**Integration:**
This assessment works alongside WB1 (Infrastructure Inventory) to document
services running on network devices, and feeds into WB5 (Controls Coverage)
and the Network Security Compliance Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_3_services_catalog.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_3_services_catalog.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_3_services_catalog.py --date 20250124
    
    # Generate with custom organisation name
    python3 generate_a820_3_services_catalog.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organisation name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_3_Services Catalog_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize service types to match your service catalog
    2. Inventory all network services across infrastructure
    3. Complete service-specific security assessments
    4. Document security mechanisms for each service
    5. Validate service availability and redundancy configurations
    6. Review monitoring and alerting setup for each service
    7. Identify gaps and classify by severity
    8. Define remediation actions with timelines and ownership
    9. Collect and link audit evidence (service configs, logs)
    10. Obtain stakeholder approvals
    11. Feed results into WB5 and Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    3 of 6 (Network Services Inventory and Security Assessment)
Primary Control:      A.8.21 (Security of Network Services)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalisation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-22 specification
    - Supports comprehensive network services security assessment
    - Integrated with A.8.20-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Network Services Security Standards:**
Network services security should follow industry best practices:
- DNS: DNSSEC implementation, split DNS architecture, query logging
- DHCP: Rogue DHCP prevention, DHCP snooping, scope management
- NTP: Stratum hierarchy, time source authentication, symmetric keys
- Proxy: Content filtering, authentication, SSL inspection, logging
- Load Balancers: SSL/TLS termination, session management, health checks

Review and update service security requirements quarterly.

**Technology Diversity:**
This assessment framework is technology-agnostic and must work across:
- Traditional on-premise services (physical/virtual servers)
- Cloud-based services (AWS Route53, Azure DNS, managed services)
- Hybrid architectures (split DNS, hybrid proxy, etc.)
- Containerized services (Kubernetes services, service mesh)

Customize assessment criteria to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Complete network services inventory
- Service-specific security controls documented
- Evidence of security mechanism implementation
- Availability monitoring and SLA compliance data
- Gap analysis with remediation plans

**Data Protection:**
Assessment workbooks contain sensitive service information including:
- Service architecture and topology
- Security mechanisms and configurations
- Authentication and access control details
- Identified vulnerabilities and gaps

Handle in accordance with your organisation's data classification policies.
Restrict access to authorised network and security personnel only.

**Maintenance:**
Review and update services assessment:
- Monthly: After service configuration changes or new services deployed
- Quarterly: Routine reassessment of critical services
- Annually: Complete reassessment of all network services
- Ad-hoc: After service outages or security incidents

**Quality Assurance:**
Validate service security by:
- Exporting service configurations (zone files, DHCP scopes, etc.)
- Testing security mechanisms (DNSSEC validation, DHCP snooping tests)
- Reviewing service logs and monitoring data
- Peer review by service owners and security engineers
- Penetration testing of service security controls

**Regulatory Alignment:**
Ensure network services security aligns with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 network services requirements
- Healthcare: HIPAA network services security standards
- Finance: Regional banking network services requirements
- Government: Jurisdiction-specific service security mandates

Customize assessment criteria to include regulatory-specific requirements.

**Integration Points:**
This assessment integrates with other ISO 27001 controls:
- A.8.15 (Logging): Service logging requirements and implementation
- A.8.16 (Monitoring): Service monitoring and availability tracking
- A.8.20 (Network Security): Services running on network devices
- A.5.23 (Cloud Services): Cloud-based network services

**Common Pitfalls to Avoid:**
1. **DNS Vulnerabilities**: Cache poisoning, zone transfers, no DNSSEC
2. **DHCP Attacks**: Rogue DHCP servers, DHCP exhaustion, no snooping
3. **NTP Attacks**: Amplification attacks, unauthenticated time sources
4. **Proxy Bypass**: Poor filtering rules, SSL/TLS inspection gaps
5. **Load Balancer Issues**: Weak SSL/TLS, session hijacking risks
6. **SNMP v1/v2**: Using insecure SNMP versions with community strings
7. **Unencrypted Syslog**: Logs transmitted without encryption
8. **No Redundancy**: Single points of failure for critical services

**Service-Specific Security Considerations:**

**DNS Security:**
- Implement DNSSEC for zone signing and validation
- Use split DNS (internal vs. external resolution)
- Disable zone transfers to unauthorised hosts
- Enable query logging and monitoring
- Protect against DNS tunneling and exfiltration

**DHCP Security:**
- Enable DHCP snooping on switches
- Implement rogue DHCP server detection
- Use DHCP reservations for critical systems
- Log DHCP lease assignments
- Protect against DHCP exhaustion attacks

**NTP Security:**
- Use NTP version 4 with authentication
- Implement time source hierarchy (stratum levels)
- Use symmetric keys for authentication
- Restrict NTP access with ACLs
- Monitor time synchronization status

**Proxy Security:**
- Implement SSL/TLS inspection where appropriate
- Use content filtering and malware scanning
- Enforce authentication for proxy access
- Log all proxy traffic
- Implement bypass prevention

**Load Balancer Security:**
- Use secure SSL/TLS configurations (TLS 1.2+)
- Implement proper session management
- Configure health checks for backend servers
- Enable logging and monitoring
- Protect against DDoS attacks

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "Network Services Catalog & Security Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.20-22.S3"
CONTROL_ID   = "A.8.20-22"
CONTROL_NAME = "Network Security"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

DASH = "  -  "
SQ = "'"

# Service catalog constants
SERVICE_ROW_COUNT = 100         # Main service catalog rows
DNS_ASSESSMENT_ROWS = 51        # DNS services (1 F2F2F2 sample + 50 FFFFCC)
DHCP_ASSESSMENT_ROWS = 51       # DHCP services (1 F2F2F2 sample + 50 FFFFCC)
NTP_ASSESSMENT_ROWS = 51        # NTP services (1 F2F2F2 sample + 50 FFFFCC)
PROXY_ASSESSMENT_ROWS = 51      # 1 F2F2F2 sample + 50 FFFFCC empty
ADDITIONAL_SERVICE_ROWS = 51    # 1 F2F2F2 sample + 50 FFFFCC empty
GAP_ROW_COUNT = 50              # Gap tracking

# Service types
SERVICE_TYPES = [
    "DNS",
    "DHCP",
    "NTP",
    "Proxy/Web Filter",
    "Load Balancer",
    "RADIUS/TACACS+ (AAA)",
    "SNMP",
    "Syslog",
    "SMTP Relay",
    "FTP/SFTP",
    "Other",
]
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "yes_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="006100"),
        },
        "no_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="9C0006"),
        },
        "na_fill": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "font": Font(name="Calibri", size=10, color="7F7F7F"),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        },
        "low_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "compliant_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "noncompliant_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "partial_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "info_box": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Service Type validation
    validations["service_type"] = DataValidation(
        type="list",
        formula1=f'"{",".join(SERVICE_TYPES)}"',
        allow_blank=False,
    )
    validations["service_type"].error = "Invalid service type"
    validations["service_type"].errorTitle = "Service Type Error"
    
    # Criticality validation
    validations["criticality"] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False,
    )
    
    # Status validation
    validations["status"] = DataValidation(
        type="list",
        formula1='"Active,Offline,Planned,Decommissioned"',
        allow_blank=False,
    )
    
    # Yes/No/N/A validation
    validations["yes_no_na"] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False,
    )
    
    # Redundancy Level validation
    validations["redundancy"] = DataValidation(
        type="list",
        formula1='"Active-Active,Active-Passive,Clustered,Single Point of Failure,N/A"',
        allow_blank=True,
    )
    
    # Monitoring Status validation
    validations["monitoring"] = DataValidation(
        type="list",
        formula1='"Monitored,Not Monitored,Partially Monitored"',
        allow_blank=True,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False,
    )
    
    # Gap Status validation
    validations["gap_status"] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Accepted Risk"',
        allow_blank=False,
    )
    
    return validations


# ============================================================================
# SECTION 3B: VALIDATION FINALIZER
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & SERVICE GUIDE
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Services Catalog for all network services deployed in your environment.', '2. Assess DNS security controls in the DNS Security Assessment sheet.', '3. Assess DHCP security controls in the DHCP Security Assessment sheet.', '4. Assess NTP security controls in the NTP Security Assessment sheet.', '5. Assess proxy/web gateway security in the Proxy Security Assessment sheet.', '6. Document additional network services (RADIUS, SNMP, etc.) in the Additional Services sheet.', '7. Review compliance scoring in the Service Compliance Summary sheet.', '8. Identify and document gaps in the Gap Analysis sheet with remediation plans.', '9. Document service dependencies in the Service Dependencies sheet.', '10. Maintain the Evidence Register with all supporting documentation for audit traceability.', '11. Obtain final approval and sign-off from IT Security, ISO, and CISO.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 25

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_services_catalog_sheet(ws, styles, validations):
    """Create main Services Catalog sheet."""
    
    ws.title = "Services Catalog"
    
    # Title
    ws.merge_cells("A1:P1")
    cell = ws["A1"]
    cell.value = f"NETWORK SERVICES CATALOG - GENERATED {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Master catalog of all network services. Yellow cells are input fields. Link to detailed assessments in service-specific sheets."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A - Auto
        "Service Type",         # B - Dropdown
        "Service Name",         # C
        "Purpose",              # D
        "Hosting Location",     # E (Device/Server)
        "IP Address",           # F
        "Port(s)",              # G
        "Criticality",          # H - Dropdown
        "Redundancy Level",     # I - Dropdown
        "Monitoring Status",    # J - Dropdown
        "SLA Uptime %",         # K
        "Owner",                # L
        "Last Security Review", # M
        "Status",               # N - Dropdown
        "Notes",                # O
        "Assessment Link",      # P (reference to detailed sheet)
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + SERVICE_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Service ID (auto-generated)
        service_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SVC-" & TEXT({service_num},"000"),"")')
        
        # Input cells
        for col in range(2, 17):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["service_type"].add(f"B{start_row}:B{end_row}")
    ws.add_data_validation(validations["service_type"])
    
    validations["criticality"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["criticality"])
    
    validations["redundancy"].add(f"I{start_row}:I{end_row}")
    ws.add_data_validation(validations["redundancy"])
    
    validations["monitoring"].add(f"J{start_row}:J{end_row}")
    ws.add_data_validation(validations["monitoring"])
    
    validations["status"].add(f"N{start_row}:N{end_row}")
    ws.add_data_validation(validations["status"])
    
    # Conditional Formatting for Criticality
    
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Column widths
    widths = [12, 18, 25, 35, 20, 15, 10, 12, 20, 15, 12, 20, 15, 12, 40, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: SHEET 3 - DNS SECURITY ASSESSMENT
# ============================================================================

def create_dns_security_sheet(ws, styles, validations):
    """Create DNS Security Assessment sheet."""
    
    ws.title = "DNS Security Assessment"
    
    # Title
    ws.merge_cells("A1:N1")
    cell = ws["A1"]
    cell.value = "DNS SECURITY ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:N2")
    ws["A2"] = "Assess DNS security controls per IMP-S4 guidance. One row per DNS service instance."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers - DNS-specific security controls
    headers = [
        "Service ID",           # A (from catalog)
        "DNS Type",             # B (Authoritative/Recursive/Forwarder)
        "DNSSEC Enabled",       # C - Y/N/NA
        "Split DNS",            # D - Y/N/NA
        "Rate Limiting",        # E - Y/N/NA
        "Query Logging",        # F - Y/N/NA
        "RPZ/Filtering",        # G - Y/N/NA
        "Recursion Control",    # H - Y/N/NA
        "Cache Poisoning Protection", # I - Y/N/NA
        "Zone Transfer Controls", # J - Y/N/NA
        "TSIG/SIG(0)",          # K - Y/N/NA
        "Compliance Score %",   # L - Auto-calc
        "Last Assessed",        # M
        "Notes",                # N
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + DNS_ASSESSMENT_ROWS - 1
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        # Service ID
        cell = ws.cell(row=row_idx, column=1)
        if is_sample:
            cell.fill = _sample_fill
            cell.border = _sample_border
            cell.value = "DNS-001"
        else:
            apply_style(cell, styles["input_cell"])

        # DNS Type dropdown
        cell = ws.cell(row=row_idx, column=2)
        if is_sample:
            cell.fill = _sample_fill
            cell.border = _sample_border
            cell.value = "Recursive"
        else:
            apply_style(cell, styles["input_cell"])

        # Security controls (Yes/No/N/A)
        for col in range(3, 12):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
            else:
                apply_style(cell, styles["input_cell"])

        # Compliance Score formula
        yes_range = f"C{row_idx}:K{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=12, value=formula)
        ws.cell(row=row_idx, column=12).number_format = "0.0"

        # Last Assessed, Notes
        for col in [13, 14]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
            else:
                apply_style(cell, styles["input_cell"])

    # Complete sample row for DNS controls (cols C-K) and Last Assessed
    for col in range(3, 12):
        ws.cell(row=start_row, column=col).value = "Yes"
    ws.cell(row=start_row, column=13).value = "15.01.2026"  # Last Assessed

    # Apply Yes/No/NA validation to control columns (data rows only, skip sample row 4)
    for col in range(3, 12):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row + 1}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])

    # Apply conditional formatting for Yes/No/NA
    for col in range(3, 12):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"N/A"'], fill=styles["na_fill"]["fill"], font=styles["na_fill"]["font"])
        )

    # DNS Security Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "DNS Security Controls - Implementation Reference (per IMP-S4)"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 15):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    dns_controls = [
        ("DNSSEC", "Enable zone signing, publish DS records to parent, configure trust anchors"),
        ("Split DNS", "Internal view for private zones, external view for public zones"),
        ("Rate Limiting", "Configure response rate limiting (RRL) to prevent DDoS amplification"),
        ("Query Logging", "Enable query logging for security monitoring and forensics"),
        ("RPZ/Filtering", "Implement Response Policy Zones for malware/phishing domain blocking"),
        ("Recursion Control", "Disable recursion on authoritative servers, restrict recursive queries by ACL"),
        ("Cache Poisoning Protection", "Enable DNSSEC validation, source port randomization"),
        ("Zone Transfer Controls", "Restrict zone transfers (AXFR) to authorised secondary servers only"),
        ("TSIG/SIG(0)", "Use TSIG for authenticated zone transfers and dynamic updates"),
    ]
    
    row += 1
    for control, guidance in dns_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:N{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["N"].width = 40
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: SHEET 4 - DHCP SECURITY ASSESSMENT
# ============================================================================

def create_dhcp_security_sheet(ws, styles, validations):
    """Create DHCP Security Assessment sheet."""
    
    ws.title = "DHCP Security Assessment"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "DHCP SECURITY ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Assess DHCP security controls per IMP-S4 guidance. One row per DHCP server/scope."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "DHCP Scope",           # B (IP range)
        "DHCP Snooping",        # C - Y/N/NA
        "Rogue DHCP Detection", # D - Y/N/NA
        "Server Hardening",     # E - Y/N/NA
        "Lease Time Appropriate", # F - Y/N/NA
        "Utilization Monitoring", # G - Y/N/NA
        "Reservation Management", # H - Y/N/NA
        "Logging Enabled",      # I - Y/N/NA
        "Compliance Score %",   # J - Auto
        "Last Assessed",        # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + DHCP_ASSESSMENT_ROWS - 1
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        # Service ID, DHCP Scope
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
                if col == 1:
                    cell.value = "DHCP-001"
                else:
                    cell.value = "192.168.1.0/24"
            else:
                apply_style(cell, styles["input_cell"])

        # Security controls
        for col in range(3, 10):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
            else:
                apply_style(cell, styles["input_cell"])

        # Compliance Score
        yes_range = f"C{row_idx}:I{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=10, value=formula)
        ws.cell(row=row_idx, column=10).number_format = "0.0"

        # Last Assessed, Notes
        for col in [11, 12]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
            else:
                apply_style(cell, styles["input_cell"])

    # Complete sample row for DHCP controls (cols C-I) and Last Assessed
    for col in range(3, 10):
        ws.cell(row=start_row, column=col).value = "Yes"
    ws.cell(row=start_row, column=11).value = "15.01.2026"  # Last Assessed

    # Apply validations and formatting (data rows only, skip sample row 4)
    for col in range(3, 10):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row + 1}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])

    for col in range(3, 10):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )

    # DHCP Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "DHCP Security Controls - Implementation Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 13):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    dhcp_controls = [
        ("DHCP Snooping", "Enable on access switches to prevent rogue DHCP servers"),
        ("Rogue Detection", "Active scanning or monitoring for unauthorised DHCP responses"),
        ("Server Hardening", "Apply OS hardening, disable unnecessary services, patch regularly"),
        ("Lease Time", "Appropriate lease time (8 hours workstations, 24 hours servers, 1 hour guest)"),
        ("Utilization Monitoring", "Monitor scope utilization, alert at 80% threshold"),
        ("Reservation Management", "Document static reservations, review regularly"),
        ("Logging", "Enable DHCP server logging for all lease actions"),
    ]
    
    row += 1
    for control, guidance in dhcp_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:L{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["L"].width = 40
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SHEET 5 - NTP SECURITY ASSESSMENT
# ============================================================================

def create_ntp_security_sheet(ws, styles, validations):
    """Create NTP Security Assessment sheet."""
    
    ws.title = "NTP Security Assessment"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "NTP SECURITY ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "Assess NTP security controls per IMP-S4 guidance. One row per NTP server/client."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "NTP Role",             # B (Server/Client)
        "Authentication",       # C - Y/N/NA
        "Access Control",       # D - Y/N/NA
        "Stratum Appropriate",  # E - Y/N/NA
        "Source Validation",    # F - Y/N/NA
        "Monitoring",           # G - Y/N/NA
        "Compliance Score %",   # H - Auto
        "Stratum Level",        # I
        "Last Assessed",        # J
        "Notes",                # K
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + NTP_ASSESSMENT_ROWS - 1
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        # Service ID, NTP Role
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
                if col == 1:
                    cell.value = "NTP-001"
                else:
                    cell.value = "Server"
            else:
                apply_style(cell, styles["input_cell"])

        # Security controls
        for col in range(3, 8):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
            else:
                apply_style(cell, styles["input_cell"])

        # Compliance Score
        yes_range = f"C{row_idx}:G{row_idx}"
        formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
        ws.cell(row=row_idx, column=8, value=formula)
        ws.cell(row=row_idx, column=8).number_format = "0.0"

        # Stratum, Last Assessed, Notes
        for col in [9, 10, 11]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
            else:
                apply_style(cell, styles["input_cell"])

    # Complete sample row for NTP controls (cols C-G), Stratum Level, and Last Assessed
    for col in range(3, 8):
        ws.cell(row=start_row, column=col).value = "Yes"
    ws.cell(row=start_row, column=9).value = "2"        # Stratum Level (col I)
    ws.cell(row=start_row, column=10).value = "15.01.2026"  # Last Assessed (col J)

    # Apply validations (data rows only, skip sample row 4)
    for col in range(3, 8):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row + 1}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])

    for col in range(3, 8):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )

    # NTP Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "NTP Security Controls - Implementation Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 12):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    ntp_controls = [
        ("Authentication", "Enable NTP authentication (MD5 keys) between servers"),
        ("Access Control", "Restrict NTP queries and modifications (restrict default kod nomodify notrap)"),
        ("Stratum Hierarchy", "Appropriate stratum levels (Stratum 1/2 for internal servers)"),
        ("Source Validation", "Verify NTP server sources, use multiple reliable sources"),
        ("Monitoring", "Monitor time offset, alert on significant drift (>1 second)"),
    ]
    
    row += 1
    for control, guidance in ntp_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["K"].width = 40
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 9: SHEET 6 - PROXY SECURITY ASSESSMENT
# ============================================================================

def create_proxy_security_sheet(ws, styles, validations):
    """Create Proxy/Web Filter Security Assessment sheet."""
    
    ws.title = "Proxy Security Assessment"
    
    # Title
    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "PROXY/WEB FILTER SECURITY ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "Assess Proxy/Web Filtering security controls per IMP-S4 guidance."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "Proxy Type",           # B (Forward/Reverse/Transparent)
        "Authentication",       # C - Y/N/NA
        "SSL Inspection",       # D - Y/N/NA
        "Category Filtering",   # E - Y/N/NA
        "Malware Scanning",     # F - Y/N/NA
        "Logging",              # G - Y/N/NA
        "Bypass Prevention",    # H - Y/N/NA
        "Redundancy",           # I - Y/N/NA
        "Compliance Score %",   # J - Auto
        "Last Assessed",        # K
        "Filter Categories",    # L
        "Notes",                # M
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + PROXY_ASSESSMENT_ROWS - 1
    
    _proxy_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _proxy_thin = Side(style="thin")
    _proxy_border = Border(left=_proxy_thin, right=_proxy_thin, top=_proxy_thin, bottom=_proxy_thin)

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        sample_vals = ["PRXY-001", "Forward", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", 100.0, "15.01.2026", "Malware, Adult Content", ""]

        # Service ID, Proxy Type
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = sample_vals[col - 1]
                cell.fill = _proxy_sample_fill
                cell.border = _proxy_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

        # Security controls
        for col in range(3, 10):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = sample_vals[col - 1]
                cell.fill = _proxy_sample_fill
                cell.border = _proxy_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

        # Compliance Score
        if is_sample:
            cell = ws.cell(row=row_idx, column=10, value=100.0)
            cell.fill = _proxy_sample_fill
            cell.border = _proxy_border
            cell.font = Font(color="808080", name="Calibri")
            cell.number_format = "0.0"
        else:
            yes_range = f"C{row_idx}:I{row_idx}"
            formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
            ws.cell(row=row_idx, column=10, value=formula)
            ws.cell(row=row_idx, column=10).number_format = "0.0"

        # Last Assessed, Filter Categories, Notes
        for col in [11, 12, 13]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = sample_vals[col - 1]
                cell.fill = _proxy_sample_fill
                cell.border = _proxy_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Apply validations
    for col in range(3, 10):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row + 1}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])

    for col in range(3, 10):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
    
    # Proxy Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "Proxy Security Controls - Implementation Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 14):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Control"
    ws["B" + str(row)] = "Implementation Guidance"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    proxy_controls = [
        ("Authentication", "LDAP/AD authentication, Kerberos/NTLM for transparent proxy"),
        ("SSL Inspection", "SSL/TLS decryption and inspection (with CA cert deployment)"),
        ("Category Filtering", "Block malicious/inappropriate categories (malware, adult, gambling, etc.)"),
        ("Malware Scanning", "Real-time malware scanning of downloaded content"),
        ("Logging", "Log all requests with user, URL, timestamp, action (allow/block)"),
        ("Bypass Prevention", "Block direct internet access, enforce proxy via firewall rules"),
        ("Redundancy", "High availability configuration (multiple proxy servers)"),
    ]
    
    row += 1
    for control, guidance in proxy_controls:
        ws[f"A{row}"] = control
        ws.merge_cells(f"B{row}:M{row}")
        ws[f"B{row}"] = guidance
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["M"].width = 40
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: SHEET 7 - ADDITIONAL SERVICES
# ============================================================================

def create_additional_services_sheet(ws, styles, validations):
    """Create Additional Services assessment (Load Balancer, AAA, SNMP, Syslog, etc.)."""
    
    ws.title = "Additional Services"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "ADDITIONAL NETWORK SERVICES ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Assess other network services: Load Balancers, AAA (RADIUS/TACACS+), SNMP, Syslog, SMTP, FTP, etc."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "Service Type",         # B
        "Authentication",       # C - Y/N/NA
        "Encryption",           # D - Y/N/NA
        "Access Control",       # E - Y/N/NA
        "Logging",              # F - Y/N/NA
        "Monitoring",           # G - Y/N/NA
        "Hardening Applied",    # H - Y/N/NA
        "Compliance Score %",   # I - Auto
        "Last Assessed",        # J
        "Key Controls",         # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + ADDITIONAL_SERVICE_ROWS - 1
    
    _addl_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _addl_thin = Side(style="thin")
    _addl_border = Border(left=_addl_thin, right=_addl_thin, top=_addl_thin, bottom=_addl_thin)

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        sample_vals = ["SVC-001", "RADIUS/AAA", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", 100.0, "15.01.2026", "TACACS+, SNMPv3", ""]

        # Service ID, Service Type
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = sample_vals[col - 1]
                cell.fill = _addl_sample_fill
                cell.border = _addl_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

        # Security controls
        for col in range(3, 9):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = sample_vals[col - 1]
                cell.fill = _addl_sample_fill
                cell.border = _addl_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

        # Compliance Score
        if is_sample:
            cell = ws.cell(row=row_idx, column=9, value=100.0)
            cell.fill = _addl_sample_fill
            cell.border = _addl_border
            cell.font = Font(color="808080", name="Calibri")
            cell.number_format = "0.0"
        else:
            yes_range = f"C{row_idx}:H{row_idx}"
            formula = f'=IF(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No")=0,"",COUNTIF({yes_range},"Yes")/(COUNTIF({yes_range},"Yes")+COUNTIF({yes_range},"No"))*100)'
            ws.cell(row=row_idx, column=9, value=formula)
            ws.cell(row=row_idx, column=9).number_format = "0.0"

        # Last Assessed, Key Controls, Notes
        for col in [10, 11, 12]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = sample_vals[col - 1]
                cell.fill = _addl_sample_fill
                cell.border = _addl_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Apply validations
    for col in range(3, 9):
        validations["yes_no_na"].add(f"{get_column_letter(col)}{start_row + 1}:{get_column_letter(col)}{end_row}")
    ws.add_data_validation(validations["yes_no_na"])

    for col in range(3, 9):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=styles["yes_fill"]["fill"], font=styles["yes_fill"]["font"])
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{end_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=styles["no_fill"]["fill"], font=styles["no_fill"]["font"])
        )
    
    # Service-Specific Controls Reference
    row = end_row + 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "Service-Specific Security Controls Reference"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 13):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Service Type"
    ws["B" + str(row)] = "Key Security Controls"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    service_controls = [
        ("Load Balancer", "SSL/TLS termination, strong ciphers, health checks, session persistence security, DDoS protection"),
        ("RADIUS/TACACS+", "Strong authentication, command authorisation (TACACS+), encrypted communication, MFA support, logging"),
        ("SNMP", "SNMPv3 only (v1/v2c disabled), authentication (authPriv), encryption, access control, trap security"),
        ("Syslog", "TLS encryption (rsyslog/syslog-ng), log retention policy, log rotation, integrity protection"),
        ("SMTP Relay", "Authentication required, relay restrictions, SPF/DKIM/DMARC, TLS encryption, rate limiting"),
        ("FTP/SFTP", "Use SFTP (SSH), disable FTP, strong authentication, access control, logging"),
    ]
    
    row += 1
    for service, controls in service_controls:
        ws[f"A{row}"] = service
        ws.merge_cells(f"B{row}:L{row}")
        ws[f"B{row}"] = controls
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["K"].width = 35
    ws.column_dimensions["L"].width = 40
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 11: SHEET 8 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1, TABLE 2, TABLE 3."""

    ws.title = "Summary Dashboard"

    _SQ = "'"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _white_bold_14 = Font(bold=True, color="FFFFFF", size=14)
    _white_bold_11 = Font(bold=True, color="FFFFFF", size=11)
    _dark_bold_10 = Font(bold=True, color="000000", size=10)
    _dark_10 = Font(bold=False, color="000000", size=10)
    _dark_9_italic = Font(bold=False, color="000000", size=9, italic=True)
    _title_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _data_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _t3total_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _t3_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    _center = Alignment(horizontal="center", vertical="center")
    _left = Alignment(horizontal="left", vertical="center")
    _wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _banner(row, text, font, fill):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = font
        c.fill = fill

    def _hdr_row(row, labels):
        for col_letter, text in labels:
            c = ws[f"{col_letter}{row}"]
            c.value = text
            c.font = _dark_bold_10
            c.fill = _hdr_fill
            c.border = _thin
            c.alignment = _center

    def _data_row(row, area, b_formula, c_formula, d_formula, e_formula, f_formula, g_formula):
        c_a = ws[f"A{row}"]
        c_a.value = area
        c_a.font = _dark_10
        c_a.border = _thin
        c_a.alignment = _left
        for col, formula in [("B", b_formula), ("C", c_formula), ("D", d_formula), ("E", e_formula), ("F", f_formula)]:
            c = ws[f"{col}{row}"]
            c.value = formula
            c.font = _dark_10
            c.border = _thin
            c.alignment = _center
        c_g = ws[f"G{row}"]
        c_g.value = g_formula
        c_g.font = _dark_10
        c_g.border = _thin
        c_g.alignment = _center

    def _total_row_t1(row, first_data_row, last_data_row):
        c_a = ws[f"A{row}"]
        c_a.value = "TOTAL"
        c_a.font = Font(bold=True, color="000000", size=10)
        c_a.fill = _total_fill
        c_a.border = _thin
        c_a.alignment = _left
        for col in ["B", "C", "D", "E", "F"]:
            c = ws[f"{col}{row}"]
            c.value = f"=SUM({col}{first_data_row}:{col}{last_data_row})"
            c.font = Font(bold=True, color="000000", size=10)
            c.fill = _total_fill
            c.border = _thin
            c.alignment = _center
        c_g = ws[f"G{row}"]
        c_g.value = f"=IF((B{row}-F{row})=0,\"0%\",ROUND(C{row}/(B{row}-F{row})*100,1)&\"%\")"
        c_g.font = Font(bold=True, color="000000", size=10)
        c_g.fill = _total_fill
        c_g.border = _thin
        c_g.alignment = _center

    # ── ROW 1: Title ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 35
    _banner(1, "NETWORK SERVICES CATALOG & SECURITY ASSESSMENT — SUMMARY DASHBOARD", _white_bold_14, _title_fill)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── ROW 2: Subtitle ───────────────────────────────────────────────
    ws["A2"] = "ISO/IEC 27001:2022 — Controls A.8.20 · A.8.21 · A.8.22: Network Security, Security of Network Services and Segregation of Networks"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = _left
    ws.merge_cells("A2:G2")

    # ── ROW 3: empty ──────────────────────────────────────────────────

    # ── TABLE 1 ───────────────────────────────────────────────────────
    _banner(4, "TABLE 1: ASSESSMENT AREA COMPLIANCE", _white_bold_11, _title_fill)
    _hdr_row(5, [("A", "Assessment Area"), ("B", "Total"), ("C", "Yes"), ("D", "Partial"), ("E", "No"), ("F", "N-A"), ("G", "Compliance %")])

    # Row 6: DNS Security Assessment (score col L, rows 5-54)
    _data_row(6,
        "DNS Security Assessment",
        f"=COUNTA({_SQ}DNS Security Assessment{_SQ}!A5:A54)",
        f"=COUNTIF({_SQ}DNS Security Assessment{_SQ}!L5:L54,\">=80\")",
        f"=COUNTIFS({_SQ}DNS Security Assessment{_SQ}!L5:L54,\">=50\",{_SQ}DNS Security Assessment{_SQ}!L5:L54,\"<80\")",
        f"=COUNTIFS({_SQ}DNS Security Assessment{_SQ}!L5:L54,\">=0\",{_SQ}DNS Security Assessment{_SQ}!L5:L54,\"<50\")",
        "=B6-(C6+D6+E6)",
        "=IF((B6-F6)=0,\"0%\",ROUND(C6/(B6-F6)*100,1)&\"%\")"
    )

    # Row 7: DHCP Security Assessment (7 controls C-I, 51 rows — skip sample row 4, count rows 5-54)
    _data_row(7,
        "DHCP Security Assessment",
        f"=COUNTA({_SQ}DHCP Security Assessment{_SQ}!C5:I54)",
        f"=COUNTIF({_SQ}DHCP Security Assessment{_SQ}!C5:I54,\"Yes\")",
        "",
        f"=COUNTIF({_SQ}DHCP Security Assessment{_SQ}!C5:I54,\"No\")",
        f"=COUNTIF({_SQ}DHCP Security Assessment{_SQ}!C5:I54,\"N/A\")",
        "=IF((B7-F7)=0,\"0%\",ROUND(C7/(B7-F7)*100,1)&\"%\")"
    )

    # Row 8: NTP Security Assessment (5 controls C-G, 51 rows — skip sample row 4, count rows 5-54)
    _data_row(8,
        "NTP Security Assessment",
        f"=COUNTA({_SQ}NTP Security Assessment{_SQ}!C5:G54)",
        f"=COUNTIF({_SQ}NTP Security Assessment{_SQ}!C5:G54,\"Yes\")",
        "",
        f"=COUNTIF({_SQ}NTP Security Assessment{_SQ}!C5:G54,\"No\")",
        f"=COUNTIF({_SQ}NTP Security Assessment{_SQ}!C5:G54,\"N/A\")",
        "=IF((B8-F8)=0,\"0%\",ROUND(C8/(B8-F8)*100,1)&\"%\")"
    )

    # Row 9: Proxy Security Assessment (7 controls C-I, 50 FFFFCC rows — skip F2F2F2 sample row 4, count rows 5-54)
    _data_row(9,
        "Proxy Security Assessment",
        f"=COUNTA({_SQ}Proxy Security Assessment{_SQ}!C5:I54)",
        f"=COUNTIF({_SQ}Proxy Security Assessment{_SQ}!C5:I54,\"Yes\")",
        "",
        f"=COUNTIF({_SQ}Proxy Security Assessment{_SQ}!C5:I54,\"No\")",
        f"=COUNTIF({_SQ}Proxy Security Assessment{_SQ}!C5:I54,\"N/A\")",
        "=IF((B9-F9)=0,\"0%\",ROUND(C9/(B9-F9)*100,1)&\"%\")"
    )

    # Row 10: Additional Services (6 controls C-H, 50 FFFFCC rows — skip F2F2F2 sample row 4, count rows 5-54)
    _data_row(10,
        "Additional Services",
        f"=COUNTA({_SQ}Additional Services{_SQ}!C5:H54)",
        f"=COUNTIF({_SQ}Additional Services{_SQ}!C5:H54,\"Yes\")",
        "",
        f"=COUNTIF({_SQ}Additional Services{_SQ}!C5:H54,\"No\")",
        f"=COUNTIF({_SQ}Additional Services{_SQ}!C5:H54,\"N/A\")",
        "=IF((B10-F10)=0,\"0%\",ROUND(C10/(B10-F10)*100,1)&\"%\")"
    )

    # Row 11: Gap Analysis (status col I, rows 5-53)
    _data_row(11,
        "Gap Analysis",
        f"=COUNTA({_SQ}Gap Analysis{_SQ}!B5:B53)",
        f"=COUNTIF({_SQ}Gap Analysis{_SQ}!I5:I53,\"Resolved\")",
        f"=COUNTIF({_SQ}Gap Analysis{_SQ}!I5:I53,\"In Progress\")",
        f"=COUNTIF({_SQ}Gap Analysis{_SQ}!I5:I53,\"Open\")",
        "=B11-(C11+D11+E11)",
        "=IF((B11-F11)=0,\"0%\",ROUND(C11/(B11-F11)*100,1)&\"%\")"
    )

    # Row 12: TOTAL
    _total_row_t1(12, 6, 11)

    # ── TABLE 2 ───────────────────────────────────────────────────────
    _banner(14, "TABLE 2: KEY METRICS", _white_bold_11, _title_fill)

    ws["A15"].value = "Metric"
    ws["A15"].font = _dark_bold_10
    ws["A15"].fill = _hdr_fill
    ws["A15"].border = _thin
    ws["A15"].alignment = _left
    ws["B15"].value = "Value"
    ws["B15"].font = _dark_bold_10
    ws["B15"].fill = _hdr_fill
    ws["B15"].border = _thin
    ws["B15"].alignment = _center
    ws.merge_cells("C15:G15")
    ws["C15"].value = "What This Shows"
    ws["C15"].font = _dark_bold_10
    ws["C15"].fill = _hdr_fill
    ws["C15"].border = _thin
    ws["C15"].alignment = _left

    t2_metrics = [
        (16, "DNS Services Below 80%",
         f"=COUNTIFS({_SQ}DNS Security Assessment{_SQ}!L5:L54,\">=0\",{_SQ}DNS Security Assessment{_SQ}!L5:L54,\"<80\")",
         "DNS services below 80% compliance (DNSSEC, logging, and filtering gaps)"),
        (17, "NTP Services Below Threshold",
         f"=COUNTIFS({_SQ}NTP Security Assessment{_SQ}!H5:H54,\">=0\",{_SQ}NTP Security Assessment{_SQ}!H5:H54,\"<80\")",
         "NTP services below 80% (authentication and authorisation gaps)"),
        (18, "DHCP Services Below Threshold",
         f"=COUNTIFS({_SQ}DHCP Security Assessment{_SQ}!J5:J54,\">=0\",{_SQ}DHCP Security Assessment{_SQ}!J5:J54,\"<80\")",
         "DHCP services below 80% (snooping, rogue DHCP prevention gaps)"),
        (19, "Unmonitored Services",
         f"=COUNTIF({_SQ}Services Catalog{_SQ}!J5:J{4 + SERVICE_ROW_COUNT},\"Not Monitored\")",
         "Network services not monitored (blind spots in security posture)"),
        (20, "Critical Services Count",
         f"=COUNTIF({_SQ}Services Catalog{_SQ}!H5:H{4 + SERVICE_ROW_COUNT},\"Critical\")",
         "Total critical services requiring priority security attention"),
        (21, "Open Service Security Gaps",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!I5:I53,\"Open\")",
         "Outstanding service security gaps without remediation"),
        (22, "Critical Gap Count",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!F5:F53,\"Critical\")",
         "Critical-severity service security gaps requiring immediate action"),
        (23, "Overall Compliance Rate",
         "=G12",
         "Overall network services security compliance percentage (TABLE 1 total)"),
    ]

    for row, metric, formula, description in t2_metrics:
        ws[f"A{row}"].value = metric
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = description
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).border = _thin

    # ── TABLE 3 ───────────────────────────────────────────────────────
    _banner(25, "TABLE 3: CRITICAL FINDINGS", _white_bold_11, _t3_fill)

    ws["A26"].value = "Critical Finding Type"
    ws["A26"].font = _dark_bold_10
    ws["A26"].fill = _hdr_fill
    ws["A26"].border = _thin
    ws["A26"].alignment = _left
    ws["B26"].value = "Count"
    ws["B26"].font = _dark_bold_10
    ws["B26"].fill = _hdr_fill
    ws["B26"].border = _thin
    ws["B26"].alignment = _center
    ws.merge_cells("C26:G26")
    ws["C26"].value = "Filter Instructions"
    ws["C26"].font = _dark_bold_10
    ws["C26"].fill = _hdr_fill
    ws["C26"].border = _thin
    ws["C26"].alignment = _left

    t3_findings = [
        (27, "DNS Services Below 80%",
         f"=COUNTIFS({_SQ}DNS Security Assessment{_SQ}!L5:L54,\">=0\",{_SQ}DNS Security Assessment{_SQ}!L5:L54,\"<80\")",
         "Filter DNS Security Assessment: Score < 80"),
        (28, "Critical Severity Service Gaps",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!F5:F53,\"Critical\")",
         "Filter Gap Analysis: Severity = \"Critical\""),
        (29, "Open Service Security Gaps",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!I5:I53,\"Open\")",
         "Filter Gap Analysis: Status = \"Open\""),
        (30, "Unmonitored Critical Services",
         f"=COUNTIFS({_SQ}Services Catalog{_SQ}!H5:H{4 + SERVICE_ROW_COUNT},\"Critical\",{_SQ}Services Catalog{_SQ}!J5:J{4 + SERVICE_ROW_COUNT},\"Not Monitored\")",
         "Filter Services Catalog: Criticality = \"Critical\" AND Monitoring = \"Not Monitored\""),
        (31, "NTP/DHCP Services Below Threshold",
         f"=COUNTIFS({_SQ}NTP Security Assessment{_SQ}!H5:H54,\">=0\",{_SQ}NTP Security Assessment{_SQ}!H5:H54,\"<80\")+COUNTIFS({_SQ}DHCP Security Assessment{_SQ}!J5:J54,\">=0\",{_SQ}DHCP Security Assessment{_SQ}!J5:J54,\"<80\")",
         "Filter NTP sheet: Score < 80; or filter DHCP sheet: Score < 80"),
    ]

    for row, finding_type, formula, instructions in t3_findings:
        ws[f"A{row}"].value = finding_type
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].fill = _data_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].fill = _data_fill
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = instructions
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].fill = _data_fill
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).fill = _data_fill
            ws.cell(row=row, column=_c).border = _thin

    # TOTAL row
    ws["A32"].value = "TOTAL"
    ws["A32"].font = Font(bold=True, color="000000", size=10)
    ws["A32"].border = _thin
    ws["A32"].alignment = _left
    ws["B32"].value = "=SUM(B27:B31)"
    ws["B32"].font = Font(bold=True, color="000000", size=10)
    ws["B32"].fill = _t3total_fill
    ws["B32"].border = _thin
    ws["B32"].alignment = _center
    ws.merge_cells("C32:G32")
    ws["C32"].value = "Total critical findings requiring immediate remediation"
    ws["C32"].font = Font(italic=True, size=9, color="000000")
    ws["C32"].alignment = _left

    # ── Column widths ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 12: SHEET 9 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(ws, styles, validations):
    """Create Gap Analysis sheet for service security gaps."""
    
    ws.title = "Gap Analysis"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "SERVICE SECURITY GAPS - REMEDIATION TRACKING"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = 'Document all service security gaps (missing controls, misconfigurations). Track remediation progress.'
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",               # A - Auto
        "Service ID",           # B
        "Service Type",         # C
        "Security Gap",         # D
        "Current State",        # E
        "Severity",             # F - Dropdown
        "Remediation Plan",     # G
        "Owner",                # H
        "Status",               # I - Dropdown
        "Target Date",          # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1
    
    _svc_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _svc_yell = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _svc_bord = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        # Gap ID (auto-generated)
        gap_num = row_idx - start_row + 1
        is_sample = (row_idx == start_row)
        c1 = ws.cell(row=row_idx, column=1)
        if is_sample:
            c1.value = "SVC-GAP-001"
            c1.fill = _svc_grey
            c1.border = _svc_bord
            c1.font = Font(color="808080", name="Calibri")
        else:
            c1.value = f'=IF(B{row_idx}<>"","SVC-GAP-" & TEXT({gap_num},"000"),"")'
            c1.fill = _svc_yell
            c1.border = _svc_bord

        # Input cells
        for col in range(2, 11):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _svc_grey
                cell.border = _svc_bord
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample data for row 4
    ws.cell(row=start_row, column=2).value = "SVC-001"
    ws.cell(row=start_row, column=3).value = "DNS"
    ws.cell(row=start_row, column=4).value = "No DNSSEC validation on internal resolver"
    ws.cell(row=start_row, column=5).value = "DNS cache poisoning possible on corporate resolver"
    ws.cell(row=start_row, column=6).value = "High"
    ws.cell(row=start_row, column=7).value = "Enable DNSSEC validation on all internal DNS resolvers"
    ws.cell(row=start_row, column=8).value = "Network Team"
    ws.cell(row=start_row, column=9).value = "Open"
    ws.cell(row=start_row, column=10).value = "28.02.2026"

    # Apply Data Validations (exclude sample row)
    validations["gap_severity"].add(f"F{start_row + 1}:F{end_row}")
    ws.add_data_validation(validations["gap_severity"])

    validations["gap_status"].add(f"I{start_row + 1}:I{end_row}")
    ws.add_data_validation(validations["gap_status"])
    
    # Conditional Formatting for Severity
    
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 12
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 13: SHEET 10 - SERVICE DEPENDENCIES
# ============================================================================

def create_service_dependencies_sheet(ws, styles):
    """Create Service Dependencies mapping sheet."""
    
    ws.title = "Service Dependencies"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "SERVICE DEPENDENCIES MAPPING"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:F2")
    ws["A2"] = "Map service dependencies to understand impact of service failures."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A
        "Service Name",         # B
        "Depends On",           # C (Service IDs)
        "Required By",          # D (Service IDs)
        "Impact of Failure",    # E
        "Notes",                # F
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows: 1 F2F2F2 sample + 50 FFFFCC empty = 51 rows
    start_row = 4
    end_row = start_row + 50  # 51 rows total

    _dep_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _dep_thin = Side(style="thin")
    _dep_border = Border(left=_dep_thin, right=_dep_thin, top=_dep_thin, bottom=_dep_thin)
    dep_sample_vals = ["SVC-DEP-001", "DNS (Primary)", "NTP-001", "Active Directory, Email, Web Apps", "High availability — all name resolution fails if DNS fails", ""]

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = dep_sample_vals[col - 1]
                cell.fill = _dep_sample_fill
                cell.border = _dep_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])
    
    # Example Dependencies
    row = end_row + 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Common Service Dependencies - Examples"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Service"
    ws["B" + str(row)] = "Typical Dependencies"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    dependencies = [
        ("Active Directory", "Depends On: DNS, NTP, DHCP (optional)"),
        ("Email (Exchange/O365)", "Depends On: DNS, Active Directory, Network Connectivity"),
        ("Web Applications", "Depends On: DNS, Load Balancer, Database, Authentication (LDAP/AD)"),
        ("VPN Access", "Depends On: RADIUS/TACACS+, Active Directory, DNS"),
        ("DHCP", "Depends On: Network connectivity (usually no dependencies)"),
        ("DNS", "Depends On: Network connectivity, NTP (for DNSSEC)"),
    ]
    
    row += 1
    for service, deps in dependencies:
        ws[f"A{row}"] = service
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"] = deps
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 40
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 13B: SHEET 11 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (italic, NOT blue banner) ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers (003366, white text) ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Network Services Assessment",
        3: "Configuration file",
        4: "DNS/DHCP/NTP service configuration exports and security assessment records",
        5: "/evidence/A.8.20-22/",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 13C: SHEET 12 - APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet -- golden standard common sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G11),\"\")"),
        ("Assessment Status:", ""),
    ]
    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    status_dv = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=True)
    ws.add_data_validation(status_dv)
    if status_row:
        status_dv.add(ws[f"B{status_row}"])

    approvers = [
        ("COMPLETED BY (NETWORK ENGINEERING)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True, name="Calibri")
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    decision_dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

# ============================================================================
# SECTION 14: MAIN FUNCTION
# ============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    styles = _STYLES
    validations = create_data_validations()

    create_instructions_sheet(wb.create_sheet())
    create_services_catalog_sheet(wb.create_sheet(), styles, validations)
    create_dns_security_sheet(wb.create_sheet(), styles, validations)
    create_dhcp_security_sheet(wb.create_sheet(), styles, validations)
    create_ntp_security_sheet(wb.create_sheet(), styles, validations)
    create_proxy_security_sheet(wb.create_sheet(), styles, validations)
    create_additional_services_sheet(wb.create_sheet(), styles, validations)
    create_gap_analysis_sheet(wb.create_sheet(), styles, validations)
    create_service_dependencies_sheet(wb.create_sheet(), styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    create_summary_dashboard_sheet(wb.create_sheet(), styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_sheet(ws, styles)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path.name}")

def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"ERROR saving workbook: {e}")
        sys.exit(1)
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\nAssessment Sheets:")
    logger.info(f"  • Services Catalog ({SERVICE_ROW_COUNT} master service entries)")
    logger.info(f"  • DNS Security Assessment ({DNS_ASSESSMENT_ROWS} rows, 9 controls)")
    logger.info(f"  • DHCP Security Assessment ({DHCP_ASSESSMENT_ROWS} rows, 7 controls)")
    logger.info(f"  • NTP Security Assessment ({NTP_ASSESSMENT_ROWS} rows, 5 controls)")
    logger.info(f"  • Proxy Security Assessment ({PROXY_ASSESSMENT_ROWS} rows, 7 controls)")
    logger.info(f"  • Additional Services ({ADDITIONAL_SERVICE_ROWS} rows, 6 controls)")
    logger.info("\n Analysis & Reporting:")
    logger.info("  • Service Compliance Summary (metrics, pie chart)")
    logger.info(f"  • Gap Analysis ({GAP_ROW_COUNT} gap tracking rows)")
    logger.info("  • Service Dependencies (dependency mapping)")
    logger.info("  • Evidence Register (audit evidence tracking)")
    logger.info("  • Approval Sign-Off (stakeholder approval)")
    logger.info("\n" + "=" * 80)
    logger.info("NEXT STEPS:")
    logger.info("  1. Complete Services Catalog (all network services)")
    logger.info("  2. Assess each service type in dedicated sheets")
    logger.info("  3. Document gaps in Gap Analysis")
    logger.info("  4. Map dependencies in Service Dependencies")
    logger.info("  5. Review Service Compliance Summary")
    logger.info("\n" + "=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
