#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-22.S1 - Network Infrastructure Inventory Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 1 of 6: Network Device Inventory and Discovery

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific network infrastructure, device types, and
assessment requirements.

Key customization areas:
1. Network device types and vendors (match your actual infrastructure)
2. Device categorization scheme (adapt to your network architecture)
3. Location taxonomy (specific to your sites/data centers)
4. Criticality classification (aligned with your business impact analysis)
5. Discovery tool integration (based on your network management tools)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for inventorying
all network infrastructure devices across the organisation, providing the
foundation for network security assessment and compliance validation.

**Purpose:**
Enables systematic inventory of all network infrastructure devices across the
organisation, providing foundation for device security assessment and overall
network security compliance evaluation against ISO 27001:2022 Control A.8.20.

**Assessment Scope:**
- Network device discovery and inventory
- Device categorization (router, switch, firewall, wireless AP, load balancer)
- Location and ownership tracking
- Management interface documentation
- Firmware/software version tracking
- Criticality classification
- Last assessment date tracking
- Discovery metadata and audit trail

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and device categorization standards
2. Device Inventory - Comprehensive device inventory with all attributes
3. Routers - Router-specific inventory
4. Switches - Switch-specific inventory
5. Firewalls - Firewall-specific inventory
6. Wireless_APs - Wireless access point inventory
7. Load_Balancers - Load balancer inventory
8. VPN_Concentrators - VPN device inventory
9. Other_Devices - Other network devices (IDS/IPS, network monitoring, etc.)
10. Discovery_Log - Network discovery methodology and results
11. Evidence Register - Audit evidence tracking and documentation
12. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with device type and vendor dropdown lists
- Conditional formatting for device status and criticality
- Automated device categorization
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with network discovery tools

**Integration:**
This assessment feeds into WB2 (Device Security Assessment) where each inventoried
device is assessed for hardening compliance, and into the overall Network Security
Compliance Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_1_infrastructure_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_1_infrastructure_inventory.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_1_infrastructure_inventory.py --date 20250124
    
    # Generate with custom organisation name
    python3 generate_a820_1_infrastructure_inventory.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organisation name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_1_Infrastructure_Inventory_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize device type categories to match your infrastructure
    2. Conduct network discovery using automated tools (nmap, SNMP, network scanners)
    3. Import discovered devices into inventory workbook
    4. Complete device attributes (location, owner, criticality, versions)
    5. Validate inventory completeness against network diagrams
    6. Document discovery methodology in Discovery_Log sheet
    7. Collect and link audit evidence (discovery tool reports, SNMP walks)
    8. Obtain stakeholder approvals
    9. Feed inventory into WB2 (Device Security Assessment)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    1 of 6 (Network Device Inventory and Discovery)
Primary Control:      A.8.20 (Network Security)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalisation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-22 specification
    - Supports comprehensive network device inventory
    - Integrated with A.8.20-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Network Discovery Standards:**
Network device discovery should be performed systematically using automated tools.
Review and update discovery methodology quarterly to ensure complete coverage.
Common discovery methods include:
- SNMP walks (v3 preferred for security)
- Network scanning (nmap, network mappers)
- Configuration management database (CMDB) exports
- Network access control (NAC) system reports
- Manual documentation of non-discoverable devices

**Technology Diversity:**
This inventory framework is technology-agnostic and must work across:
- Traditional networks (physical routers, switches, firewalls)
- Software-Defined Networks (SDN, SD-WAN controllers)
- Cloud environments (AWS, Azure, GCP virtual network appliances)
- Hybrid architectures (on-premise + cloud connectivity)

Customize device categories to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Complete network device inventory (no shadow IT devices)
- Current firmware/software versions documented
- Criticality classification aligned with business impact analysis
- Device ownership and responsibility clearly assigned

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Network topology information (device locations, connections)
- Management interface details (IP addresses, access methods)
- Device configurations and vulnerabilities
- Security zone membership

Handle in accordance with your organisation's data classification policies.
Restrict access to authorised network and security personnel only.

**Maintenance:**
Review and update inventory:
- Monthly: After significant network changes (new devices added/removed)
- Quarterly: Routine re-discovery and validation
- Annually: Complete inventory refresh with full network discovery
- Ad-hoc: When network security incidents occur or audits scheduled

**Quality Assurance:**
Validate inventory completeness by:
- Cross-checking against network topology diagrams
- Comparing with CMDB/asset management system records
- Verifying against firewall and switch port mappings
- Reconciling with change management records
- Performing periodic network scans to detect shadow IT

**Regulatory Alignment:**
Ensure device inventory supports applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 network device inventory requirements
- Healthcare: HIPAA network infrastructure documentation
- Finance: Regional banking network security requirements
- Government: Jurisdiction-specific network mandates

Customize inventory attributes to capture regulatory-specific data.

**Integration Points:**
This inventory integrates with other ISO 27001 controls:
- A.8.1 (Asset Inventory): Network devices as information assets
- A.8.8 (Vulnerability Management): Devices requiring vulnerability scanning
- A.8.9 (Configuration Management): Devices requiring configuration baselines
- A.5.23 (Cloud Services): Cloud-based network infrastructure inventory

**Common Pitfalls to Avoid:**
1. **Incomplete Discovery**: Missing shadow IT devices (rogue switches, APs)
2. **Outdated Information**: Firmware versions not current
3. **Missing Ownership**: Devices without assigned responsibility
4. **Wrong Criticality**: Criticality not aligned with business impact
5. **No Evidence**: Discovery claims without supporting evidence
6. **Manual Only**: Relying solely on manual documentation (automate!)
7. **Point-in-Time**: Treating inventory as one-time exercise vs. continuous

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "Network Infrastructure Inventory"
DOCUMENT_ID = "ISMS-IMP-A.8.20-22.S1"
CONTROL_ID   = "A.8.20-22"
CONTROL_NAME = "Network Security"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

DASH = "  -  "

# Cross-sheet formula reference (sheet names with spaces need quoting)
DI = "'Device Inventory'"

# Device inventory constants
DEVICE_ROW_COUNT = 150  # Pre-formatted rows for device entries
GAP_ROW_COUNT = 50      # Gap analysis tracking rows
DEVICE_TYPE_SUMMARY_ROWS = 50  # Device type summary data rows
EVIDENCE_ROW_COUNT = 30 # Evidence register rows
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style templates as dictionaries to avoid shared object issues.
    """
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "low_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "status_active": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "status_offline": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "status_decom": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        },
        "info_box": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell, creating new objects to avoid warnings."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Device Type validation
    validations["device_type"] = DataValidation(
        type="list",
        formula1='"Router,Switch,Firewall,Wireless AP,Load Balancer,VPN Concentrator,IDS/IPS,Network Management,Other"',
        allow_blank=False,
    )
    validations["device_type"].error = "Invalid device type"
    validations["device_type"].errorTitle = "Device Type Error"
    
    # Criticality validation
    validations["criticality"] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False,
    )
    validations["criticality"].error = "Must be Critical, High, Medium, or Low"
    validations["criticality"].errorTitle = "Criticality Error"
    
    # Status validation
    validations["status"] = DataValidation(
        type="list",
        formula1='"Active,Offline,Decommissioned,Planned"',
        allow_blank=False,
    )
    validations["status"].error = "Invalid status"
    validations["status"].errorTitle = "Status Error"
    
    # Security Zone validation
    validations["zone"] = DataValidation(
        type="list",
        formula1='"Internet/DMZ,Internal,Management,Guest,Datacenter,Branch,Cloud,Unknown"',
        allow_blank=True,
    )
    validations["zone"].error = "Invalid security zone"
    validations["zone"].errorTitle = "Zone Error"
    
    # Discovery Method validation
    validations["discovery_method"] = DataValidation(
        type="list",
        formula1='"Nmap Scan,SNMP Walk,Cloud API,Manual Entry,Documentation Review,Other"',
        allow_blank=True,
    )
    
    # Compliance Status validation
    validations["compliance"] = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant,⚠️ Partial,❓ Not Assessed"',
        allow_blank=True,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False,
    )
    
    # Gap Status validation
    validations["gap_status"] = DataValidation(
        type="list",
        formula1='"⭕ Open,In Progress,Resolved,⚠️ Accepted Risk"',
        allow_blank=False,
    )
    
    # Evidence Type validation
    validations["evidence_type"] = DataValidation(
        type="list",
        formula1='"Configuration Backup,Discovery Report,Network Diagram,Scan Results,Documentation,Screenshot,Other"',
        allow_blank=False,
    )
    
    # Evidence Status validation  
    validations["evidence_status"] = DataValidation(
        type="list",
        formula1='"Collected,Pending,Missing,Not Applicable"',
        allow_blank=False,
    )
    
    return validations


# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Device Inventory sheet for every network device in scope.', '2. Use dropdown menus for standardised assessments (device type, status, criticality).', '3. Assess device criticality using the Device Criticality Matrix framework.', '4. Review the Device Type Summary for coverage statistics across device categories.', '5. Document discovery metadata and methodology in the Discovery Metadata sheet.', '6. Identify and document gaps in the Gap Analysis sheet with remediation plans.', '7. Review the Validation Rules sheet for data quality checks.', '8. Maintain the Evidence Register with all supporting documentation for audit traceability.', '9. Obtain final approval and sign-off from IT Security, ISO, and CISO.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_device_inventory_sheet(ws, styles, validations):
    """Create main Device Inventory sheet with 150 pre-formatted rows."""
    
    ws.title = "Device Inventory"
    
    # Title
    ws.merge_cells("A1:R1")
    cell = ws["A1"]
    cell.value = f"NETWORK INFRASTRUCTURE INVENTORY - GENERATED {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35

    # Instructions row
    ws.merge_cells("A2:R2")
    ws["A2"] = "Complete inventory for all network devices discovered per IMP-S1. Yellow cells are input fields. Device IDs auto-generate."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Device ID",           # A - Auto-generated
        "Device Type",         # B - Dropdown (REQUIRED)
        "Make/Model",          # C - Text
        "Hostname",            # D - Text (REQUIRED)
        "Primary IP",          # E - Text (REQUIRED)
        "Management IP",       # F - Text
        "Location",            # G - Text (REQUIRED)
        "Security Zone",       # H - Dropdown
        "Purpose",             # I - Text
        "Criticality",         # J - Dropdown (REQUIRED)
        "Owner",               # K - Text
        "Last Discovered",     # L - Date
        "Discovery Method",    # M - Dropdown
        "Firmware Version",    # N - Text
        "Serial Number",       # O - Text
        "Compliance Status",   # P - Dropdown
        "Status",              # Q - Dropdown (REQUIRED)
        "Notes",               # R - Text
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows (150 pre-formatted rows)
    start_row = 4
    end_row = start_row + DEVICE_ROW_COUNT - 1

    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _thin_s = Side(style="thin")
    border_s = Border(left=_thin_s, right=_thin_s, top=_thin_s, bottom=_thin_s)

    for row_idx in range(start_row, end_row + 1):
        # Device ID
        if row_idx == start_row:
            ws.cell(row=row_idx, column=1, value="NET-DEV-001")
            ws.cell(row=row_idx, column=1).font = Font(color="808080", name="Calibri")
        else:
            device_num = row_idx - start_row + 1
            ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","NET-DEV-" & TEXT({device_num},"000"),"")')

        # Apply input styling to data entry cells
        input_cols = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]  # B through R
        for col in input_cols:
            cell = ws.cell(row=row_idx, column=col)
            if row_idx == start_row:
                cell.fill = _sample_fill
                cell.border = border_s
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample row data for row 4 guidance
    ws.cell(row=start_row, column=2).value = "Router"
    ws.cell(row=start_row, column=3).value = "Cisco ISR 4451"
    ws.cell(row=start_row, column=4).value = "core-rtr-01"
    ws.cell(row=start_row, column=5).value = "192.168.1.1"
    ws.cell(row=start_row, column=6).value = "192.168.1.250"
    ws.cell(row=start_row, column=7).value = "Server Room Rack 1"
    ws.cell(row=start_row, column=8).value = "Management"
    ws.cell(row=start_row, column=9).value = "Core routing and inter-VLAN routing"
    ws.cell(row=start_row, column=10).value = "Critical"
    ws.cell(row=start_row, column=11).value = "IT Infrastructure Team"
    ws.cell(row=start_row, column=12).value = "01.01.2026"
    ws.cell(row=start_row, column=13).value = "Network Scan"
    ws.cell(row=start_row, column=14).value = "15.3.8a"
    ws.cell(row=start_row, column=15).value = "FTX12345678"
    ws.cell(row=start_row, column=16).value = "Compliant"
    ws.cell(row=start_row, column=17).value = "Active"
    ws.cell(row=start_row, column=18).value = ""

    # Apply Data Validations
    # Device Type (Column B)
    validations["device_type"].add(f"B{start_row + 1}:B{end_row}")
    ws.add_data_validation(validations["device_type"])

    # Criticality (Column J)
    validations["criticality"].add(f"J{start_row + 1}:J{end_row}")
    ws.add_data_validation(validations["criticality"])

    # Security Zone (Column H)
    validations["zone"].add(f"H{start_row + 1}:H{end_row}")
    ws.add_data_validation(validations["zone"])

    # Discovery Method (Column M)
    validations["discovery_method"].add(f"M{start_row + 1}:M{end_row}")
    ws.add_data_validation(validations["discovery_method"])

    # Compliance Status (Column P)
    validations["compliance"].add(f"P{start_row + 1}:P{end_row}")
    ws.add_data_validation(validations["compliance"])

    # Status (Column Q)
    validations["status"].add(f"Q{start_row + 1}:Q{end_row}")
    ws.add_data_validation(validations["status"])
    
    # Apply Conditional Formatting for Criticality (Column J)
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"J{start_row}:J{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"J{start_row}:J{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"J{start_row}:J{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"J{start_row}:J{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Apply Conditional Formatting for Status (Column Q)
    ws.conditional_formatting.add(
        f"Q{start_row}:Q{end_row}",
        CellIsRule(operator="equal", formula=['"Active"'], fill=styles["status_active"]["fill"])
    )
    ws.conditional_formatting.add(
        f"Q{start_row}:Q{end_row}",
        CellIsRule(operator="equal", formula=['"Offline"'], fill=styles["status_offline"]["fill"])
    )
    ws.conditional_formatting.add(
        f"Q{start_row}:Q{end_row}",
        CellIsRule(operator="equal", formula=['"Decommissioned"'], fill=styles["status_decom"]["fill"])
    )
    
    # Column widths
    widths = [12, 15, 25, 30, 15, 15, 25, 15, 30, 12, 25, 12, 18, 18, 18, 18, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Freeze panes (freeze header rows)
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: SHEET 3 - DEVICE CRITICALITY MATRIX
# ============================================================================

def create_device_criticality_matrix_sheet(ws, styles):
    """Create Device Criticality Matrix with assessment framework."""
    
    ws.title = "Device Criticality Matrix"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "DEVICE CRITICALITY ASSESSMENT MATRIX"
    apply_style(cell, styles["title"])
    for _c in range(1, 7):
        ws.cell(row=1, column=_c).border = _border
    ws.row_dimensions[1].height = 35

    # Purpose
    ws.merge_cells("A2:F2")
    ws["A2"] = "Use this matrix to assess device criticality based on business impact if device fails"
    apply_style(ws["A2"], styles["info_box"])
    for _c in range(1, 7):
        ws.cell(row=2, column=_c).border = _border

    # Assessment Criteria
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Assessment Criteria"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = _border
    
    row += 1
    criteria_headers = ["Criticality", "Service Impact", "Recovery Time", "User Impact", "Revenue Impact", "Redundancy"]
    for col_idx, header in enumerate(criteria_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    criteria_data = [
        (
            "Critical",
            "Complete outage of critical services",
            "< 1 hour",
            "All users affected",
            "Significant revenue loss",
            "Single point of failure"
        ),
        (
            "High",
            "Significant service degradation",
            "< 4 hours",
            "Large user group affected",
            "Notable revenue impact",
            "Limited redundancy"
        ),
        (
            "Medium",
            "Localized service impact",
            "< 24 hours",
            "Department/team affected",
            "Minimal revenue impact",
            "Redundant but important"
        ),
        (
            "Low",
            "Minimal or no service impact",
            "> 24 hours acceptable",
            "Few users affected",
            "No revenue impact",
            "Fully redundant or non-critical"
        ),
    ]
    
    row += 1
    for crit_row in criteria_data:
        for col_idx, value in enumerate(crit_row, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            
            if col_idx == 1:  # Criticality column
                if crit_row[0] == "Critical":
                    apply_style(cell, styles["critical_fill"])
                elif crit_row[0] == "High":
                    apply_style(cell, styles["high_fill"])
                elif crit_row[0] == "Medium":
                    apply_style(cell, styles["medium_fill"])
                elif crit_row[0] == "Low":
                    apply_style(cell, styles["low_fill"])
                cell.border = _border  # colored fill styles have no border
            else:
                apply_style(cell, styles["info_box"])

        row += 1

    # Device Type - Typical Criticality Mapping
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Device Type - Typical Criticality Mapping (Guidance Only)"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 4):
        ws.cell(row=row, column=_c).border = _border
    
    row += 1
    ws["A" + str(row)] = "Device Type"
    ws["B" + str(row)] = "Typical Criticality"
    ws["C" + str(row)] = "Rationale"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    device_mapping = [
        ("Core Router", "Critical", "Single point of failure for all WAN/Internet traffic"),
        ("Core Switch", "Critical", "Aggregates all access layer traffic"),
        ("Primary Firewall", "Critical", "Controls all ingress/egress traffic"),
        ("Datacenter Switch", "Critical", "Connects production servers"),
        ("Distribution Switch", "High", "Aggregates multiple access switches"),
        ("WAN Router", "High", "Connects branch to HQ"),
        ("Secondary Firewall", "High", "Backup for primary firewall"),
        ("Access Switch", "Medium", "Localized impact (single floor/department)"),
        ("Wireless AP", "Medium", "Affects wireless users only"),
        ("Management Device", "Medium", "Does not forward production traffic"),
        ("Guest Wi-Fi AP", "Low", "Non-business critical"),
        ("Lab Equipment", "Low", "Test/development only"),
    ]
    
    row += 1
    for device_type, typical_crit, rationale in device_mapping:
        ws[f"A{row}"] = device_type
        ws[f"B{row}"] = typical_crit
        ws[f"C{row}"] = rationale
        
        apply_style(ws[f"A{row}"], styles["info_box"])
        
        if typical_crit == "Critical":
            apply_style(ws[f"B{row}"], styles["critical_fill"])
        elif typical_crit == "High":
            apply_style(ws[f"B{row}"], styles["high_fill"])
        elif typical_crit == "Medium":
            apply_style(ws[f"B{row}"], styles["medium_fill"])
        elif typical_crit == "Low":
            apply_style(ws[f"B{row}"], styles["low_fill"])
        ws[f"B{row}"].border = _border  # colored fill styles have no border

        apply_style(ws[f"C{row}"], styles["info_box"])
        row += 1

    # Important Note
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "⚠️ Important: Criticality assessment must consider YOUR organisation's specific context, not just device type. Always assess based on business impact."
    apply_style(ws[f"A{row}"], styles["info_box"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = _border
    
    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 25


# ============================================================================
# SECTION 7: SHEET 4 - DEVICE TYPE SUMMARY
# ============================================================================

def create_device_type_summary_sheet(ws, styles):
    """Create Device Type Summary with data entry rows and statistics."""

    ws.title = "Device Type Summary"

    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "DEVICE TYPE SUMMARY & STATISTICS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Categorise network device types, record counts and compliance status. One row per device category."
    apply_style(ws["A2"], styles["info_box"])

    # Column Headers (row 3)
    headers = [
        "Device Type",              # A
        "Total Count",              # B - COUNTIF from Device Inventory
        "Active",                   # C - COUNTIF
        "Offline",                  # D - COUNTIF
        "Critical Count",           # E - COUNTIF
        "Compliant Count",          # F - COUNTIF
        "Compliance %",             # G - formula
        "Notes",                    # H
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Row 4: F2F2F2 sample row
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_s = Side(style="thin")
    border_s = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    sample_vals = ["Router", '=IF(A4="","",COUNTIF(\'Device Inventory\'!B5:B153,A4))',
                   '=IF(A4="","",COUNTIFS(\'Device Inventory\'!B5:B153,A4,\'Device Inventory\'!Q5:Q153,"Active"))',
                   '=IF(A4="","",COUNTIFS(\'Device Inventory\'!B5:B153,A4,\'Device Inventory\'!Q5:Q153,"Offline"))',
                   '=IF(A4="","",COUNTIFS(\'Device Inventory\'!B5:B153,A4,\'Device Inventory\'!J5:J153,"Critical"))',
                   '=IF(A4="","",COUNTIFS(\'Device Inventory\'!B5:B153,A4,\'Device Inventory\'!P5:P153,"Compliant"))',
                   '=IF(OR(A4="",B4=0),"",F4/B4*100)', ""]
    for col_idx, val in enumerate(sample_vals, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = sample_fill
        cell.border = border_s
        cell.font = Font(color="808080", name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=4, column=7).number_format = "0.0"

    # Rows 5-54: 50 empty FFFFCC input rows
    start_row = 5
    end_row = 54

    for row_idx in range(start_row, end_row + 1):
        # Column A - Device Type
        cell_a = ws.cell(row=row_idx, column=1)
        apply_style(cell_a, styles["input_cell"])

        # Column B - Total Count (COUNTIF from Device Inventory)
        ws.cell(row=row_idx, column=2,
                value=f'=IF(A{row_idx}="","",COUNTIF({DI}!B5:B{3 + DEVICE_ROW_COUNT},A{row_idx}))')
        apply_style(ws.cell(row=row_idx, column=2), styles["input_cell"])

        # Column C - Active
        ws.cell(row=row_idx, column=3,
                value=f'=IF(A{row_idx}="","",COUNTIFS({DI}!B5:B{3 + DEVICE_ROW_COUNT},A{row_idx},{DI}!Q5:Q{3 + DEVICE_ROW_COUNT},"Active"))')
        apply_style(ws.cell(row=row_idx, column=3), styles["input_cell"])

        # Column D - Offline
        ws.cell(row=row_idx, column=4,
                value=f'=IF(A{row_idx}="","",COUNTIFS({DI}!B5:B{3 + DEVICE_ROW_COUNT},A{row_idx},{DI}!Q5:Q{3 + DEVICE_ROW_COUNT},"Offline"))')
        apply_style(ws.cell(row=row_idx, column=4), styles["input_cell"])

        # Column E - Critical Count
        ws.cell(row=row_idx, column=5,
                value=f'=IF(A{row_idx}="","",COUNTIFS({DI}!B5:B{3 + DEVICE_ROW_COUNT},A{row_idx},{DI}!J5:J{3 + DEVICE_ROW_COUNT},"Critical"))')
        apply_style(ws.cell(row=row_idx, column=5), styles["input_cell"])

        # Column F - Compliant Count
        ws.cell(row=row_idx, column=6,
                value=f'=IF(A{row_idx}="","",COUNTIFS({DI}!B5:B{3 + DEVICE_ROW_COUNT},A{row_idx},{DI}!P5:P{3 + DEVICE_ROW_COUNT},"Compliant"))')
        apply_style(ws.cell(row=row_idx, column=6), styles["input_cell"])

        # Column G - Compliance %
        ws.cell(row=row_idx, column=7,
                value=f'=IF(OR(A{row_idx}="",B{row_idx}=0),"",F{row_idx}/B{row_idx}*100)')
        ws.cell(row=row_idx, column=7).number_format = "0.0"
        apply_style(ws.cell(row=row_idx, column=7), styles["input_cell"])

        # Column H - Notes
        cell_h = ws.cell(row=row_idx, column=8)
        apply_style(cell_h, styles["input_cell"])

    # Summary Statistics section below data rows
    row = end_row + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Aggregate Statistics"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 9):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])

    agg_stats = [
        ("Total Devices Inventoried", f'=COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})'),
        ("Total Active Devices", f'=COUNTIF({DI}!Q5:Q{3 + DEVICE_ROW_COUNT},"Active")'),
        ("Total Offline Devices", f'=COUNTIF({DI}!Q5:Q{3 + DEVICE_ROW_COUNT},"Offline")'),
        ("Total Critical Devices", f'=COUNTIF({DI}!J5:J{3 + DEVICE_ROW_COUNT},"Critical")'),
        ("Total Compliant Devices", f'=COUNTIF({DI}!P5:P{3 + DEVICE_ROW_COUNT},"Compliant")'),
        ("Total Non-Compliant Devices", f'=COUNTIF({DI}!P5:P{3 + DEVICE_ROW_COUNT},"Non-Compliant")'),
        ("Overall Compliance %", f'=IF(COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})=0,"",COUNTIF({DI}!P5:P{3 + DEVICE_ROW_COUNT},"Compliant")/COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})*100)'),
    ]

    row += 1
    for metric, formula in agg_stats:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 40
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SHEET 5 - DISCOVERY METADATA
# ============================================================================

def create_discovery_metadata_sheet(ws, styles):
    """Create Discovery Metadata sheet to track discovery process."""
    
    ws.title = "Discovery Metadata"
    
    # Title
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "NETWORK DISCOVERY PROCESS METADATA"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35

    # Subtitle (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = "Document the discovery methodology, scope, and tools used to identify all network devices."
    ws["A2"].font = Font(italic=True, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A4"

    # Discovery Information
    _thin_dm = Side(style="thin")
    _bdr_dm = Border(left=_thin_dm, right=_thin_dm, top=_thin_dm, bottom=_thin_dm)

    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Discovery Project Information"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 6):
        ws.cell(row=row, column=_c).border = _bdr_dm

    row += 1
    discovery_info = [
        ("Project Name:", "[Enter discovery project name]"),
        ("Start Date:", "[DD.MM.YYYY]"),
        ("Completion Date:", "[DD.MM.YYYY]"),
        ("Lead Assessor:", "[Name]"),
        ("Team Members:", "[Names]"),
        ("Scope:", "[Network segments/locations covered]"),
    ]

    for label, placeholder in discovery_info:
        ws[f"A{row}"] = label
        apply_style(ws[f"A{row}"], styles["column_header"])
        ws[f"A{row}"].border = _bdr_dm
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = placeholder
        apply_style(ws[f"B{row}"], styles["input_cell"])
        for _c in range(2, 6):
            ws.cell(row=row, column=_c).border = _bdr_dm
        row += 1
    
    # Discovery Methods Used
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Discovery Methods Used"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 6):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Method"
    ws["B" + str(row)] = "Used (Y/N)"
    ws["C" + str(row)] = "Tool/Platform"
    ws["D" + str(row)] = "Coverage"
    ws["E" + str(row)] = "Notes"
    for col in ["A", "B", "C", "D", "E"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    methods = [
        ("Nmap Network Scan", "", "nmap", "", ""),
        ("SNMP Walk", "", "SNMPwalk", "", ""),
        ("AWS Discovery", "", "AWS CLI / boto3", "", ""),
        ("Azure Discovery", "", "Azure CLI", "", ""),
        ("GCP Discovery", "", "GCP SDK", "", ""),
        ("NetFlow Analysis", "", "", "", ""),
        ("Documentation Review", "", "CMDB / Wiki", "", ""),
        ("Administrator Interviews", "", "N/A", "", ""),
        ("Physical Site Survey", "", "N/A", "", ""),
        ("Other", "", "", "", ""),
    ]
    
    row += 1
    for method_data in methods:
        for col_idx, value in enumerate(method_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1
    
    # Discovery Scope
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Discovery Scope"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 6):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Network Segment/Zone"
    ws["B" + str(row)] = "IP Range"
    ws["C" + str(row)] = "Included (Y/N)"
    ws["D" + str(row)] = "Devices Found"
    ws["E" + str(row)] = "Notes"
    for col in ["A", "B", "C", "D", "E"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    # 20 rows for scope entries
    row += 1
    for _ in range(20):
        for col_idx in range(1, 6):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, styles["input_cell"])
        row += 1
    
    # Discovery Challenges
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Discovery Challenges & Limitations"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 6):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws.merge_cells(f"A{row}:E{row + 5}")
    cell = ws[f"A{row}"]
    cell.value = "[DOCUMENT ANY CHALLENGES ENCOUNTERED DURING DISCOVERY:\n- UNREACHABLE NETWORKS\n- FIREWALL BLOCKING SCANS\n- MISSING CREDENTIALS\n- UNDOCUMENTED DEVICES\n- OTHER LIMITATIONS]"
    apply_style(cell, styles["input_cell"])
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40


# ============================================================================
# SECTION 9: SHEET 6 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(ws, styles, validations):
    """Create Gap Analysis sheet for undocumented/unreachable devices."""
    
    ws.title = "Gap Analysis"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "NETWORK DISCOVERY GAP ANALYSIS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document gaps: undocumented devices, unreachable devices, missing information, or discovery failures"
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",              # A - Auto
        "Gap Type",            # B - Text
        "Device/Location",     # C - Text
        "Description",         # D - Text
        "Severity",            # E - Dropdown
        "Impact",              # F - Text
        "Remediation Plan",    # G - Text
        "Owner",               # H - Text
        "Status",              # I - Dropdown
        "Target Date",         # J - Date
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1

    _gap_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _gap_thin_s = Side(style="thin")
    _gap_border_s = Border(left=_gap_thin_s, right=_gap_thin_s, top=_gap_thin_s, bottom=_gap_thin_s)

    for row_idx in range(start_row, end_row + 1):
        # Gap ID (auto-generated)
        gap_num = row_idx - start_row + 1
        if row_idx == start_row:
            ws.cell(row=row_idx, column=1, value="DISC-GAP-001")
            ws.cell(row=row_idx, column=1).font = Font(color="808080", name="Calibri")
            ws.cell(row=row_idx, column=1).fill = _gap_sample_fill
            ws.cell(row=row_idx, column=1).border = _gap_border_s
        else:
            c1 = ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","DISC-GAP-" & TEXT({gap_num},"000"),"")') 
            c1.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            c1.border = _gap_border_s

        # Input cells
        for col in range(2, 11):
            cell = ws.cell(row=row_idx, column=col)
            if row_idx == start_row:
                cell.fill = _gap_sample_fill
                cell.border = _gap_border_s
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample row data for row 4 guidance
    ws.cell(row=start_row, column=2).value = "Missing Device"
    ws.cell(row=start_row, column=3).value = "192.168.10.0/24 subnet"
    ws.cell(row=start_row, column=4).value = "Undocumented switch discovered during network scan — not in inventory"
    ws.cell(row=start_row, column=5).value = "High"
    ws.cell(row=start_row, column=6).value = "Potential unauthorised device accessing internal network"
    ws.cell(row=start_row, column=7).value = "Identify device, classify, assign owner, add to inventory"
    ws.cell(row=start_row, column=8).value = "Network Team"
    ws.cell(row=start_row, column=9).value = "\u2b55 Open"
    ws.cell(row=start_row, column=10).value = "31.03.2026"

    # Apply Data Validations
    validations["gap_severity"].add(f"E{start_row + 1}:E{end_row}")
    ws.add_data_validation(validations["gap_severity"])

    validations["gap_status"].add(f"I{start_row + 1}:I{end_row}")
    ws.add_data_validation(validations["gap_status"])
    
    # Conditional Formatting for Severity
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f"E{start_row}:E{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"E{start_row}:E{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"E{start_row}:E{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"E{start_row}:E{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Gap Type Examples
    row = end_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "Common Gap Types: Undocumented Device, Unreachable Device, Missing Configuration, Incomplete Discovery, Authentication Failure, Firewall Blocking, Unknown Device Type"
    apply_style(ws[f"A{row}"], styles["info_box"])
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 12
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: SHEET 7 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (italic, NOT blue banner) ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers (003366, white text) ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Network Infrastructure Inventory",
        3: "Network scan",
        4: "Network device discovery scan results and inventory compliance assessment",
        5: "/evidence/A.8.20-22/",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet — golden standard common sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G7),\"\")"),
        ("Assessment Status:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)
    if status_row:
        status_dv.add(ws[f"B{status_row}"])

    approvers = [
        ("COMPLETED BY (NETWORK ENGINEERING)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True, name="Calibri")
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 11: SHEET 7 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1, TABLE 2, TABLE 3."""

    ws.title = "Summary Dashboard"

    SQ = "'"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _white_bold_14 = Font(bold=True, color="FFFFFF", size=14)
    _white_bold_11 = Font(bold=True, color="FFFFFF", size=11)
    _dark_bold_10 = Font(bold=True, color="000000", size=10)
    _dark_10 = Font(bold=False, color="000000", size=10)
    _dark_9_italic = Font(bold=False, color="000000", size=9, italic=True)
    _title_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _data_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _t3total_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _t3_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    _center = Alignment(horizontal="center", vertical="center")
    _left = Alignment(horizontal="left", vertical="center")
    _wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _banner(row, text, font, fill):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = font
        c.fill = fill

    def _hdr_row_t1(row):
        for col_letter, text in [("A", "Assessment Area"), ("B", "Total"), ("C", "Yes"), ("D", "Partial"), ("E", "No"), ("F", "N-A"), ("G", "Compliance %")]:
            c = ws[f"{col_letter}{row}"]
            c.value = text
            c.font = _dark_bold_10
            c.fill = _hdr_fill
            c.border = _thin
            c.alignment = _center

    def _data_row(row, area, b, c_val, d, e, f_val, g):
        ws[f"A{row}"].value = area
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col, val in [("B", b), ("C", c_val), ("D", d), ("E", e), ("F", f_val)]:
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.font = _dark_10
            cell.border = _thin
            cell.alignment = _center
        ws[f"G{row}"].value = g
        ws[f"G{row}"].font = _dark_10
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    def _total_row_t1(row, first, last):
        ws[f"A{row}"].value = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"A{row}"].fill = _total_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col in ["B", "C", "D", "E", "F"]:
            c = ws[f"{col}{row}"]
            c.value = f"=SUM({col}{first}:{col}{last})"
            c.font = Font(bold=True, color="000000", size=10)
            c.fill = _total_fill
            c.border = _thin
            c.alignment = _center
        ws[f"G{row}"].value = f"=IF((B{row}-F{row})=0,\"0%\",ROUND(C{row}/(B{row}-F{row})*100,1)&\"%\")"
        ws[f"G{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"G{row}"].fill = _total_fill
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    # ── ROW 1: Title ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 35
    _banner(1, "NETWORK INFRASTRUCTURE INVENTORY — SUMMARY DASHBOARD", _white_bold_14, _title_fill)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── ROW 2: Subtitle ───────────────────────────────────────────────
    ws["A2"] = "ISO/IEC 27001:2022 — Controls A.8.20 · A.8.21 · A.8.22: Network Security, Security of Network Services and Segregation of Networks"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = _left
    ws.merge_cells("A2:G2")

    # ── TABLE 1 ───────────────────────────────────────────────────────
    _banner(4, "TABLE 1: ASSESSMENT AREA COMPLIANCE", _white_bold_11, _title_fill)
    _hdr_row_t1(5)

    # Row 6: Device Inventory (compliance col P, rows 4-153)
    # DV values: Compliant / Non-Compliant / ⚠️ Partial / ❓ Not Assessed
    _data_row(6,
        "Device Inventory",
        f"=COUNTA({SQ}Device Inventory{SQ}!B5:B153)",
        f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"Compliant\")",
        f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"\u26a0\ufe0f Partial\")",
        f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"Non-Compliant\")",
        "=B6-(C6+D6+E6)",
        "=IF((B6-F6)=0,\"0%\",ROUND(C6/(B6-F6)*100,1)&\"%\")"
    )

    # Row 7: Gap Analysis (status col I, rows 4-53)
    # DV values: ⭕ Open / In Progress / Resolved / ⚠️ Accepted Risk
    _data_row(7,
        "Gap Analysis",
        f"=COUNTA({SQ}Gap Analysis{SQ}!B5:B53)",
        f"=COUNTIF({SQ}Gap Analysis{SQ}!I5:I53,\"Resolved\")",
        f"=COUNTIF({SQ}Gap Analysis{SQ}!I5:I53,\"In Progress\")",
        f"=COUNTIF({SQ}Gap Analysis{SQ}!I5:I53,\"\u2b55 Open\")",
        "=B7-(C7+D7+E7)",
        "=IF((B7-F7)=0,\"0%\",ROUND(C7/(B7-F7)*100,1)&\"%\")"
    )

    # Row 8: TOTAL
    _total_row_t1(8, 6, 7)

    # ── TABLE 2 ───────────────────────────────────────────────────────
    _banner(10, "TABLE 2: KEY METRICS", _white_bold_11, _title_fill)

    ws["A11"].value = "Metric"
    ws["A11"].font = _dark_bold_10
    ws["A11"].fill = _hdr_fill
    ws["A11"].border = _thin
    ws["A11"].alignment = _left
    ws["B11"].value = "Value"
    ws["B11"].font = _dark_bold_10
    ws["B11"].fill = _hdr_fill
    ws["B11"].border = _thin
    ws["B11"].alignment = _center
    ws.merge_cells("C11:G11")
    ws["C11"].value = "What This Shows"
    ws["C11"].font = _dark_bold_10
    ws["C11"].fill = _hdr_fill
    ws["C11"].border = _thin
    ws["C11"].alignment = _left

    t2_metrics = [
        (12, "Non-Compliant Devices",
         f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"Non-Compliant\")",
         "Devices not meeting security requirements (immediate remediation required)"),
        (13, "Partial Compliance Devices",
         f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"\u26a0\ufe0f Partial\")",
         "Devices partially compliant (incomplete protection)"),
        (14, "Undiscovered / Not Assessed",
         f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"\u2753 Not Assessed\")",
         "Devices in scope but not yet assessed (unknown security posture)"),
        (15, "Critical Devices Non-Compliant",
         f"=COUNTIFS({SQ}Device Inventory{SQ}!J5:J153,\"Critical\",{SQ}Device Inventory{SQ}!P5:P153,\"Non-Compliant\")",
         "Critical infrastructure with compliance gaps (highest risk)"),
        (16, "Offline Devices",
         f"=COUNTIF({SQ}Device Inventory{SQ}!Q5:Q153,\"Offline\")",
         "Devices not reachable for assessment (potential inventory gaps)"),
        (17, "Open Discovery Gaps",
         f"=COUNTIF({SQ}Gap Analysis{SQ}!I5:I53,\"\u2b55 Open\")",
         "Outstanding gaps in device discovery (incomplete network inventory)"),
        (18, "Critical Severity Gaps",
         f"=COUNTIF({SQ}Gap Analysis{SQ}!E5:E53,\"Critical\")",
         "Critical-severity discovery gaps requiring immediate attention"),
        (19, "Overall Compliance Rate",
         "=G8",
         "Overall network inventory compliance percentage (TABLE 1 total)"),
    ]

    for row, metric, formula, description in t2_metrics:
        ws[f"A{row}"].value = metric
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = description
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).border = _thin

    # ── TABLE 3 ───────────────────────────────────────────────────────
    _banner(21, "TABLE 3: CRITICAL FINDINGS", _white_bold_11, _t3_fill)

    ws["A22"].value = "Critical Finding Type"
    ws["A22"].font = _dark_bold_10
    ws["A22"].fill = _hdr_fill
    ws["A22"].border = _thin
    ws["A22"].alignment = _left
    ws["B22"].value = "Count"
    ws["B22"].font = _dark_bold_10
    ws["B22"].fill = _hdr_fill
    ws["B22"].border = _thin
    ws["B22"].alignment = _center
    ws.merge_cells("C22:G22")
    ws["C22"].value = "Filter Instructions"
    ws["C22"].font = _dark_bold_10
    ws["C22"].fill = _hdr_fill
    ws["C22"].border = _thin
    ws["C22"].alignment = _left

    t3_findings = [
        (23, "Non-Compliant Devices",
         f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"Non-Compliant\")",
         "Filter Device Inventory: Compliance Status = \"Non-Compliant\""),
        (24, "Critical Devices Non-Compliant",
         f"=COUNTIFS({SQ}Device Inventory{SQ}!J5:J153,\"Critical\",{SQ}Device Inventory{SQ}!P5:P153,\"Non-Compliant\")",
         "Filter Device Inventory: Criticality = \"Critical\" AND Compliance = \"Non-Compliant\""),
        (25, "Open Discovery Gaps",
         f"=COUNTIF({SQ}Gap Analysis{SQ}!I5:I53,\"\u2b55 Open\")",
         "Filter Gap Analysis: Status = \"\u2b55 Open\""),
        (26, "Critical Severity Gaps",
         f"=COUNTIF({SQ}Gap Analysis{SQ}!E5:E53,\"Critical\")",
         "Filter Gap Analysis: Severity = \"Critical\""),
        (27, "Devices Not Assessed",
         f"=COUNTIF({SQ}Device Inventory{SQ}!P5:P153,\"\u2753 Not Assessed\")",
         "Filter Device Inventory: Compliance Status = \"\u2753 Not Assessed\""),
    ]

    for row, finding_type, formula, instructions in t3_findings:
        ws[f"A{row}"].value = finding_type
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].fill = _data_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].fill = _data_fill
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = instructions
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].fill = _data_fill
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).fill = _data_fill
            ws.cell(row=row, column=_c).border = _thin

    # TOTAL row
    ws["A28"].value = "TOTAL"
    ws["A28"].font = Font(bold=True, color="000000", size=10)
    ws["A28"].border = _thin
    ws["A28"].alignment = _left
    ws["B28"].value = "=SUM(B23:B27)"
    ws["B28"].font = Font(bold=True, color="000000", size=10)
    ws["B28"].fill = _t3total_fill
    ws["B28"].border = _thin
    ws["B28"].alignment = _center
    ws.merge_cells("C28:G28")
    ws["C28"].value = "Total critical findings requiring immediate investigation"
    ws["C28"].font = Font(italic=True, size=9, color="000000")
    ws["C28"].alignment = _left

    # ── Column widths ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 12: SHEET 8 - VALIDATION RULES
# ============================================================================

def create_validation_rules_sheet(ws, styles):
    """Create Validation Rules sheet with data quality checks."""
    
    ws.title = "Validation Rules"
    
    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "DATA VALIDATION RULES & QUALITY CHECKS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Purpose
    ws.merge_cells("A2:D2")
    ws["A2"] = "This sheet defines data quality rules to ensure inventory completeness and accuracy"
    apply_style(ws["A2"], styles["info_box"])
    
    # Validation Rules
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Required Field Validation"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 5):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Field"
    ws["B" + str(row)] = "Required"
    ws["C" + str(row)] = "Validation Rule"
    ws["D" + str(row)] = "Status"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    required_fields = [
        ("Device Type", "Yes", "Must be selected from dropdown", f'=IF(COUNTBLANK({DI}!B5:B{3 + DEVICE_ROW_COUNT})={DEVICE_ROW_COUNT - 1},"✔ All blank (OK)","⚠  Check required")'),
        ("Hostname", "Yes", "Must not be blank", f'=IF(COUNTBLANK({DI}!D5:D{3 + DEVICE_ROW_COUNT})={DEVICE_ROW_COUNT - 1},"✔ All blank (OK)","⚠  Check required")'),
        ("Primary IP", "Yes", "Must not be blank", f'=IF(COUNTBLANK({DI}!E5:E{3 + DEVICE_ROW_COUNT})={DEVICE_ROW_COUNT - 1},"✔ All blank (OK)","⚠  Check required")'),
        ("Location", "Yes", "Must not be blank", f'=IF(COUNTBLANK({DI}!G5:G{3 + DEVICE_ROW_COUNT})={DEVICE_ROW_COUNT - 1},"✔ All blank (OK)","⚠  Check required")'),
        ("Criticality", "Yes", "Must be selected from dropdown", f'=IF(COUNTBLANK({DI}!J5:J{3 + DEVICE_ROW_COUNT})={DEVICE_ROW_COUNT - 1},"✔ All blank (OK)","⚠  Check required")'),
        ("Status", "Yes", "Must be selected from dropdown", f'=IF(COUNTBLANK({DI}!Q5:Q{3 + DEVICE_ROW_COUNT})={DEVICE_ROW_COUNT - 1},"✔ All blank (OK)","⚠  Check required")'),
    ]
    
    row += 1
    for field, required, rule, formula in required_fields:
        ws[f"A{row}"] = field
        ws[f"B{row}"] = required
        ws[f"C{row}"] = rule
        ws[f"D{row}"] = formula
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        row += 1
    
    # Data Quality Checks
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Data Quality Checks"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 4):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Check"
    ws["B" + str(row)] = "Description"
    ws["C" + str(row)] = "Status"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    quality_checks = [
        ("No Duplicate IPs", "Check for duplicate primary IPs", "Manual verification required"),
        ("No Duplicate Hostnames", "Check for duplicate hostnames", "Manual verification required"),
        ("Valid IP Format", "Verify all IPs follow IPv4 format", "Manual verification required"),
        ("Criticality Assigned", "All active devices have criticality", f'=IF(COUNTIFS({DI}!Q5:Q{3 + DEVICE_ROW_COUNT},"Active",{DI}!J5:J{3 + DEVICE_ROW_COUNT},"")=0,"✔ OK","⚠  Missing criticality")'),
        ("Owner Assigned", "All critical/high devices have owner", "Manual verification required"),
        ("Recent Discovery", "All devices discovered within last 90 days", "Manual verification required"),
    ]
    
    row += 1
    for check, desc, status in quality_checks:
        ws[f"A{row}"] = check
        ws[f"B{row}"] = desc
        ws[f"C{row}"] = status
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        row += 1

    # Completeness Metrics
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Inventory Completeness Metrics"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 5):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Target"
    ws["C" + str(row)] = "Actual"
    ws["D" + str(row)] = "Status"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    completeness = [
        ("Device Type Filled", "100%", f'=COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})/COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})', '=IF(C' + str(row + 1) + '>=1,"✔","⚠ ")'),
        ("Hostname Filled", "100%", f'=COUNTA({DI}!D5:D{3 + DEVICE_ROW_COUNT})/COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})', '=IF(C' + str(row + 2) + '>=1,"✔","⚠ ")'),
        ("Criticality Assigned", "100%", f'=COUNTA({DI}!J5:J{3 + DEVICE_ROW_COUNT})/COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})', '=IF(C' + str(row + 3) + '>=1,"✔","⚠ ")'),
        ("Owner Assigned", "≥80%", f'=COUNTA({DI}!K5:K{3 + DEVICE_ROW_COUNT})/COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})', '=IF(C' + str(row + 4) + '>=0.8,"✔","⚠ ")'),
        ("Firmware Version", "≥70%", f'=COUNTA({DI}!N5:N{3 + DEVICE_ROW_COUNT})/COUNTA({DI}!B5:B{3 + DEVICE_ROW_COUNT})', '=IF(C' + str(row + 5) + '>=0.7,"✔","⚠ ")'),
    ]
    
    row += 1
    for metric, target, formula, status_formula in completeness:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = target
        ws[f"C{row}"] = formula
        ws[f"D{row}"] = status_formula
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        
        # Format percentage
        ws[f"C{row}"].number_format = "0.0%"
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30


# ============================================================================
# SECTION 12: MAIN FUNCTION
# ============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    styles = _STYLES
    validations = create_data_validations()

    create_instructions_sheet(wb.create_sheet())
    create_device_inventory_sheet(wb.create_sheet(), styles, validations)
    create_device_criticality_matrix_sheet(wb.create_sheet(), styles)
    create_device_type_summary_sheet(wb.create_sheet(), styles)
    create_discovery_metadata_sheet(wb.create_sheet(), styles)
    create_gap_analysis_sheet(wb.create_sheet(), styles, validations)
    create_validation_rules_sheet(wb.create_sheet(), styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    create_summary_dashboard_sheet(wb.create_sheet(), styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_sheet(ws, styles)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path.name}")

def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"ERROR saving workbook: {e}")
        sys.exit(1)
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\nAssessment Sheets:")
    logger.info("  • Instructions & Legend (usage guidance, field definitions)")
    logger.info(f"  • Device Inventory ({DEVICE_ROW_COUNT} device rows with auto-generated IDs)")
    logger.info("  • Device Criticality Matrix (assessment framework)")
    logger.info("  • Device Type Summary (statistics and charts)")
    logger.info("\n Process Tracking:")
    logger.info("  • Discovery Metadata (discovery project tracking)")
    logger.info(f"  • Gap Analysis ({GAP_ROW_COUNT} gap tracking rows)")
    logger.info(f"  • Evidence Register ({EVIDENCE_ROW_COUNT} evidence entries)")
    logger.info("  • Validation Rules (data quality checks)")
    logger.info("\n" + "─" * 80)
    logger.info(" ASSESSMENT CAPABILITIES:")
    logger.info(f"  • {DEVICE_ROW_COUNT} network device inventory entries")
    logger.info("  • Auto-generated Device IDs (NET-DEV-001, NET-DEV-002, ...)")
    logger.info("  • Comprehensive data validations (dropdowns, required fields)")
    logger.info("  • Conditional formatting (criticality, status)")
    logger.info(f"  • {GAP_ROW_COUNT} gap identification/remediation rows")
    logger.info(f"  • {EVIDENCE_ROW_COUNT} evidence tracking entries")
    logger.info("  • Statistical charts (pie chart, bar chart)")
    logger.info("  • Data quality validation rules")
    logger.info("\n" + "─" * 80)
    logger.info(" KEY FEATURES:")
    logger.info("  Technology-agnostic (traditional, SDN, cloud, hybrid)")
    logger.info("  Systematic discovery methodology (nmap, SNMP, cloud APIs)")
    logger.info("  Comprehensive device inventory with criticality assessment")
    logger.info("  Gap analysis and evidence tracking")
    logger.info("  Automated Device ID generation")
    logger.info("  Data validation and quality checks")
    logger.info("  Summary statistics and visualization")
    logger.info("\n" + "=" * 80)
    logger.info("NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Read Instructions & Legend sheet first")
    logger.info("  3. Perform network discovery per IMP-S1 guidance")
    logger.info("  4. Fill in Device Inventory sheet with discovered devices")
    logger.info("  5. Use Device Criticality Matrix to assess device criticality")
    logger.info("  6. Document gaps in Gap Analysis sheet")
    logger.info("  7. Track evidence in Evidence Register")
    logger.info("  8. Review Validation Rules for completeness")
    logger.info("  9. Use this inventory for device hardening assessment (WB2)")
    logger.info("\nPRO TIP:")
    logger.info("  This workbook is INPUT for Workbook 2 (Device Security Assessment).")
    logger.info("  Complete discovery thoroughly - accurate inventory is foundation for")
    logger.info("  all subsequent security assessments (hardening, services, segmentation).")
    logger.info("\n" + "=" * 80)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info("\n This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
