#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-22.S5 - Security Controls Coverage Matrix Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 5 of 6: Unified Network Security Controls Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific network security architecture, control framework,
and assessment requirements.

Key customization areas:
1. Controls mapping structure (match your security control framework)
2. Security zone coverage requirements (based on your architecture)
3. Integration with other controls (A.8.15, A.8.16, etc.)
4. Coverage gap criteria (aligned with your risk assessment)
5. Evidence completeness thresholds (adapted to your audit requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook that provides
a unified view of network security controls coverage across devices, services,
and segmentation architecture, consolidating findings from WB1-4.

**Purpose:**
Enables comprehensive mapping of network security controls across devices, services,
and segmentation architecture, providing unified view of network security posture
across all three ISO 27001:2022 controls (A.8.20, A.8.21, A.8.22).

**Assessment Scope:**
- Unified controls mapping (devices → services → segmentation)
- Security zone coverage analysis
- Control effectiveness per zone
- Coverage gap identification
- Redundancy and defense-in-depth analysis
- Integration with logging (A.8.15) and monitoring (A.8.16)
- Evidence completeness tracking
- Cross-control dependency mapping

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and controls framework
2. Controls_Summary - Overall network security controls dashboard
3. Device_Controls_Coverage - Device security controls per zone
4. Service_Controls_Coverage - Network services controls per zone
5. Segmentation_Controls_Coverage - Segmentation controls per zone
6. Zone_Controls_Matrix - Unified controls matrix by security zone
7. Defense_in_Depth - Defense-in-depth analysis (layered controls)
8. Redundancy_Analysis - Control redundancy and single points of failure
9. Integration_A815_Logging - Integration with A.8.15 (Logging)
10. Integration_A816_Monitoring - Integration with A.8.16 (Monitoring)
11. Gap Analysis - Controls coverage gaps and remediation
12. Evidence_Completeness - Evidence completeness tracking
13. Cross_Control_Dependencies - Dependencies with other ISO controls
14. Evidence_Register - Audit evidence tracking and documentation
15. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with control framework dropdown lists
- Conditional formatting for coverage status (full/partial/none)
- Automated coverage gap identification
- Protected formulas with unprotected input cells
- Defense-in-depth visualization
- Single point of failure identification
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with WB1-4 for data consolidation

**Integration:**
This assessment consolidates findings from WB1 (Infrastructure Inventory),
WB2 (Device Security), WB3 (Network Services), and WB4 (Segmentation Matrix)
to provide unified controls coverage view. Feeds into Network Security
Compliance Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_5_controls_coverage.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_5_controls_coverage.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_5_controls_coverage.py --date 20250124
    
    # Generate with custom organisation name
    python3 generate_a820_5_controls_coverage.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organisation name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_5_Controls_Coverage_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Ensure WB1-4 are completed (prerequisite data)
    2. Review and customize controls framework to match your standards
    3. Map device controls to security zones (from WB2)
    4. Map service controls to security zones (from WB3)
    5. Map segmentation controls to zone boundaries (from WB4)
    6. Analyze defense-in-depth across zones
    7. Identify single points of failure and redundancy gaps
    8. Assess integration with A.8.15 (Logging) and A.8.16 (Monitoring)
    9. Document coverage gaps and remediation requirements
    10. Verify evidence completeness across all controls
    11. Collect and link audit evidence
    12. Obtain stakeholder approvals
    13. Feed results into Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    5 of 6 (Unified Network Security Controls Assessment)
Primary Control:      A.8.20-22 (Unified Network Security Framework)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalisation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-22 specification
    - Supports comprehensive unified controls coverage assessment
    - Integrated with A.8.20-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Controls Coverage Philosophy:**
This workbook provides the "big picture" view of network security controls,
answering critical questions:
- Are all security zones adequately protected?
- Do we have defense-in-depth (layered controls)?
- Where are single points of failure?
- Are controls integrated (logging, monitoring)?
- Where are coverage gaps?

Think holistically across all three controls (A.8.20, A.8.21, A.8.22).

**Technology Diversity:**
This assessment framework is technology-agnostic and must work across:
- Traditional networks (physical infrastructure)
- Software-Defined Networks (SDN, SD-WAN)
- Cloud environments (AWS, Azure, GCP)
- Hybrid architectures (on-premise + cloud)

Customize assessment criteria to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Unified view of network security controls across all zones
- Evidence of defense-in-depth strategy
- Identification of coverage gaps with remediation plans
- Integration with logging (A.8.15) and monitoring (A.8.16) documented
- Evidence completeness across all three controls

**Data Protection:**
Assessment workbooks contain comprehensive security architecture including:
- Complete network security controls mapping
- Defense-in-depth strategy and implementation
- Single points of failure and vulnerabilities
- Coverage gaps and remediation priorities

Handle in accordance with your organisation's data classification policies.
Restrict access to authorised security architects and senior management only.

**Maintenance:**
Review and update controls coverage:
- Monthly: After significant network architecture changes
- Quarterly: Routine reassessment of controls effectiveness
- Annually: Complete controls coverage review
- Ad-hoc: After security incidents or penetration test findings

**Quality Assurance:**
Validate controls coverage by:
- Cross-checking with WB1-4 data (ensure consistency)
- Verifying defense-in-depth implementation per zone
- Testing control effectiveness (not just presence)
- Peer review by security architects
- Red team assessment of controls effectiveness

**Regulatory Alignment:**
Ensure controls coverage aligns with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 defense-in-depth requirements
- Healthcare: HIPAA security controls coverage
- Finance: Regional banking controls requirements
- Government: Jurisdiction-specific control frameworks

Customize assessment criteria to include regulatory-specific requirements.

**Integration Points:**
This assessment integrates multiple ISO 27001 controls:
- A.8.20 (Network Security): Device-level controls
- A.8.21 (Network Services): Service-level controls
- A.8.22 (Segregation): Segmentation controls
- A.8.15 (Logging): Logging integration across all controls
- A.8.16 (Monitoring): Monitoring integration across all controls
- A.8.8 (Vulnerability Management): Vulnerability coverage
- A.8.9 (Configuration Management): Configuration controls

**Common Pitfalls to Avoid:**
1. **Checkbox Coverage**: Claiming controls exist without effectiveness validation
2. **No Defense-in-Depth**: Single layer of protection (no redundancy)
3. **Inconsistent Coverage**: Some zones well-protected, others neglected
4. **Single Points of Failure**: Critical controls without redundancy
5. **No Integration**: Controls operate in silos (no logging, no monitoring)
6. **Evidence Gaps**: Controls documented but no evidence of implementation
7. **Static Assessment**: Not updating coverage as architecture evolves
8. **Cloud Blindspot**: Forgetting cloud-based network security controls

**Defense-in-Depth Analysis:**

Effective defense-in-depth requires multiple layers of controls:

**Layer 1: Perimeter Controls**
- Firewalls at network boundaries
- IDS/IPS for threat detection
- DDoS protection

**Layer 2: Network Segmentation**
- Security zones with controlled boundaries
- Inter-zone traffic filtering
- Microsegmentation for critical assets

**Layer 3: Device Hardening**
- Secure device configurations
- Authentication and access control
- Encrypted management protocols

**Layer 4: Service Security**
- Service-specific security controls
- Service monitoring and availability
- Secure service configurations

**Layer 5: Monitoring and Logging**
- Comprehensive logging (A.8.15)
- Real-time monitoring (A.8.16)
- SIEM integration and alerting

**Single Point of Failure Analysis:**

Identify and remediate single points of failure:
- Single firewall protecting critical zone (needs redundancy)
- Single DNS server (needs redundant servers)
- Single VPN concentrator (needs high availability)
- Single network path (needs redundant paths)
- Single administrator with all privileges (needs role separation)

**Coverage Gap Prioritization:**

Gaps should be prioritized by:
1. **Critical**: Gaps in critical zones (no protection or failed controls)
2. **High**: Gaps in important zones or incomplete protection
3. **Medium**: Partial coverage or non-critical zone gaps
4. **Low**: Enhancement opportunities or minor gaps

**Evidence Completeness Tracking:**

For each control, track evidence:
- Configuration exports (device configs, firewall rules)
- Test results (segmentation testing, penetration tests)
- Monitoring data (logs, alerts, dashboards)
- Documentation (topology diagrams, procedures)
- Approval records (change approvals, risk acceptances)

**Cross-Control Dependencies:**

Document dependencies between network security and other controls:
- A.5.1 (Policies): Network security policies
- A.5.14 (Information Transfer): Network-based transfers
- A.5.23 (Cloud Services): Cloud network security
- A.8.1 (Asset Inventory): Network devices as assets
- A.8.8 (Vulnerability Management): Network vulnerability scanning
- A.8.9 (Configuration Management): Network configuration baselines

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, RadarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "Network Security Controls Coverage Matrix"
DOCUMENT_ID = "ISMS-IMP-A.8.20-22.S5"
CONTROL_ID   = "A.8.20-22"
CONTROL_NAME = "Network Security"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
DASH = "  -  "

OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Assessment constants
ZONE_ROW_COUNT = 51             # Zones for coverage matrix (1 F2F2F2 sample + 50 FFFFCC)
CONTROL_COLUMNS = 25            # Number of security controls
DEVICE_MAPPING_ROWS = 51        # Device-to-control mapping (1 F2F2F2 sample + 50 FFFFCC)
SERVICE_MAPPING_ROWS = 51       # 1 F2F2F2 sample + 50 FFFFCC empty
GAP_ROW_COUNT = 40              # Coverage gaps
DEFENSE_LAYER_COUNT = 51        # 1 F2F2F2 sample + 50 FFFFCC empty

# Security Control Categories (aligned with A.8.20, A.8.21, A.8.22)
CONTROL_CATEGORIES = {
    "A.8.20 - Network Security": [
        "Perimeter Firewall",
        "Network Device Hardening",
        "Network Monitoring",
        "Network Logging",
        "Wireless Security",
        "Remote Access VPN",
        "Network Access Control (NAC)",
    ],
    "A.8.21 - Network Services": [
        "DNS Security (DNSSEC)",
        "DHCP Security",
        "NTP Security",
        "Proxy/Web Filtering",
        "Load Balancer Security",
        "AAA Services (RADIUS/TACACS+)",
        "SNMP Security",
        "Syslog Security",
    ],
    "A.8.22 - Network Segmentation": [
        "Zone Segmentation",
        "VLAN Segmentation",
        "Firewall Inter-Zone Rules",
        "ACL Enforcement",
        "Segmentation Testing",
        "Microsegmentation",
    ],
}
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "covered_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="006100"),
        },
        "not_covered_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="9C0006"),
        },
        "partial_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "low_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "info_box": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Coverage Status validation
    validations["coverage_status"] = DataValidation(
        type="list",
        formula1='"Covered,Not Covered,Partially Covered,N/A"',
        allow_blank=False,
    )
    
    # Effectiveness Rating validation
    validations["effectiveness"] = DataValidation(
        type="list",
        formula1='"Effective,Partially Effective,Ineffective,Not Assessed"',
        allow_blank=False,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False,
    )
    
    # Control Category validation
    validations["control_category"] = DataValidation(
        type="list",
        formula1='"A.8.20 - Network Security,A.8.21 - Network Services,A.8.22 - Network Segmentation"',
        allow_blank=False,
    )
    
    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & INTEGRATION GUIDE
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Review the Controls Coverage Matrix to understand the full scope of the assessment.', '2. Map network device controls to security zones using data from WB1 and WB2.', '3. Map network service controls to security zones using data from WB3.', '4. Assess control effectiveness for each security zone.', '5. Validate defense-in-depth across all critical zones.', '6. Identify coverage gaps and single points of failure.', '7. Verify integration with logging (A.8.15) and monitoring (A.8.16).', '8. Document all evidence in the Evidence Register tab.', '9. Obtain required approvals in the Approval Sign-Off tab.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_controls_coverage_matrix_sheet(ws, styles, validations):
    """Create master Controls Coverage Matrix (zones × controls)."""
    
    ws.title = "Controls Coverage Matrix"
    
    # Title
    title_cols = 2 + CONTROL_COLUMNS
    ws.merge_cells(f"A1:{get_column_letter(title_cols)}1")
    cell = ws["A1"]
    cell.value = f"SECURITY CONTROLS COVERAGE MATRIX - GENERATED {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells(f"A2:{get_column_letter(title_cols)}2")
    ws["A2"] = "Matrix: Rows = Security Zones | Columns = Security Controls | Cell = Coverage Status (Covered/Not Covered/Partial)"
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers - Zone info + Controls
    row = 3
    ws.cell(row=row, column=1, value="Zone ID")
    ws.cell(row=row, column=2, value="Zone Name")
    
    col = 3
    # Flatten control categories into single list
    all_controls = []
    for category, controls in CONTROL_CATEGORIES.items():
        all_controls.extend(controls)
    
    for control in all_controls[:CONTROL_COLUMNS]:
        ws.cell(row=row, column=col, value=control)
        col += 1
    
    # Style headers
    for c in range(1, title_cols + 1):
        cell = ws.cell(row=row, column=c)
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[get_column_letter(c)].width = 16
    
    # Data rows
    start_row = 4
    end_row = start_row + ZONE_ROW_COUNT - 1
    
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        # Zone ID, Zone Name
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
                if col == 1:
                    cell.value = "ZONE-001"
                else:
                    cell.value = "Internet DMZ"
            else:
                apply_style(cell, styles["input_cell"])

        # Control coverage cells (dropdown)
        for col in range(3, 3 + CONTROL_COLUMNS):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
            else:
                apply_style(cell, styles["input_cell"])

    # Complete sample row — fill control columns C-AA with "Covered" values
    for col in range(3, 28):  # cols C through AA (columns 3-27)
        cell = ws.cell(row=start_row, column=col)
        if cell.value is None or cell.value == "":
            cell.value = "Covered"

    # Apply Data Validations to control columns (data rows only, skip sample row 4)
    for col in range(3, 3 + CONTROL_COLUMNS):
        col_letter = get_column_letter(col)
        validations["coverage_status"].add(f"{col_letter}{start_row + 1}:{col_letter}{end_row}")
    ws.add_data_validation(validations["coverage_status"])

    # Conditional Formatting for Coverage Status
    for col in range(3, 3 + CONTROL_COLUMNS):
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}{start_row + 1}:{col_letter}{end_row}"
        
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Covered"'], fill=styles["covered_fill"]["fill"], font=styles["covered_fill"]["font"])
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Not Covered"'], fill=styles["not_covered_fill"]["fill"], font=styles["not_covered_fill"]["font"])
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Partially Covered"'], fill=styles["partial_fill"]["fill"], font=styles["partial_fill"]["font"])
        )
    
    # Coverage Summary (right side)
    summary_col = 3 + CONTROL_COLUMNS
    ws.cell(row=3, column=summary_col, value="Coverage %")
    apply_style(ws.cell(row=3, column=summary_col), styles["column_header"])
    ws.column_dimensions[get_column_letter(summary_col)].width = 12
    
    # Coverage percentage formula for each zone
    for row_idx in range(start_row, end_row + 1):
        coverage_range = f"{get_column_letter(3)}{row_idx}:{get_column_letter(2 + CONTROL_COLUMNS)}{row_idx}"
        formula = f'=IF(COUNTA({coverage_range})=0,"",COUNTIF({coverage_range},"Covered")/COUNTA({coverage_range})*100)'
        ws.cell(row=row_idx, column=summary_col, value=formula)
        ws.cell(row=row_idx, column=summary_col).number_format = "0.0"

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: SHEET 3 - ZONE CONTROL ASSESSMENT
# ============================================================================

def create_zone_control_assessment_sheet(ws, styles, validations):
    """Create Zone Control Assessment - effectiveness per zone."""
    
    ws.title = "Zone Control Assessment"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "ZONE CONTROL ASSESSMENT - EFFECTIVENESS BY ZONE"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "Assess overall control effectiveness for each security zone."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Zone ID",                  # A
        "Zone Name",                # B
        "Risk Level",               # C (Critical/High/Medium/Low)
        "Controls Required",        # D (count)
        "Controls Implemented",     # E (count)
        "Control Coverage %",       # F - Auto from matrix
        "Effectiveness Rating",     # G - Dropdown
        "Gaps Identified",          # H (count)
        "Last Assessment",          # I
        "Assessed By",              # J
        "Notes",                    # K
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + ZONE_ROW_COUNT - 1
    
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        for col in range(1, 12):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
                if col == 1:
                    cell.value = "ZONE-001"
                elif col == 2:
                    cell.value = "Internet DMZ"
            else:
                apply_style(cell, styles["input_cell"])

    ws.cell(row=start_row, column=3).value = "High"
    ws.cell(row=start_row, column=4).value = 12
    ws.cell(row=start_row, column=5).value = 10
    ws.cell(row=start_row, column=6).value = "83%"
    ws.cell(row=start_row, column=7).value = "Partially Effective"
    ws.cell(row=start_row, column=8).value = 2
    ws.cell(row=start_row, column=9).value = "15.01.2026"
    ws.cell(row=start_row, column=10).value = "Network Security Team"
    ws.cell(row=start_row, column=11).value = ""

    # Apply Data Validations (data rows only, skip sample row 4)
    validations["effectiveness"].add(f"G{start_row + 1}:G{end_row}")
    ws.add_data_validation(validations["effectiveness"])

    # Conditional Formatting for Effectiveness
    ws.conditional_formatting.add(
        f"G{start_row + 1}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Effective"'], fill=styles["covered_fill"]["fill"], font=styles["covered_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"G{start_row + 1}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Ineffective"'], fill=styles["not_covered_fill"]["fill"], font=styles["not_covered_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"G{start_row + 1}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Partially Effective"'], fill=styles["partial_fill"]["fill"], font=styles["partial_fill"]["font"])
    )
    
    # Assessment Guidelines
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "Control Effectiveness Assessment Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 12):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    guidelines = [
        "Effective: All required controls implemented, tested, and functioning correctly",
        "Partially Effective: Most controls implemented, some gaps or testing incomplete",
        "Ineffective: Significant control gaps, controls not functioning, or not tested",
        "Not Assessed: Zone has not yet been assessed for control effectiveness",
    ]
    
    for guideline in guidelines:
        ws.merge_cells(f"A{row}:K{row}")
        ws[f"A{row}"] = guideline
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [12, 20, 12, 15, 15, 15, 18, 12, 12, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: SHEET 4 - DEVICE CONTROL MAPPING
# ============================================================================

def create_device_control_mapping_sheet(ws, styles):
    """Create Device Control Mapping - which devices implement which controls."""
    
    ws.title = "Device Control Mapping"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "DEVICE-TO-CONTROL MAPPING (FROM WB1 + WB2)"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "Map which network devices implement which security controls. Data from WB1 (Inventory) + WB2 (Hardening)."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Device ID",            # A (from WB1)
        "Device Type",          # B
        "Hostname",             # C
        "Security Zone",        # D
        "Controls Provided",    # E (list: Firewall, Monitoring, etc.)
        "Control Category",     # F (A.8.20/21/22)
        "Hardening Score",      # G (from WB2)
        "Effectiveness",        # H
        "Last Assessed",        # I
        "Notes",                # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + DEVICE_MAPPING_ROWS - 1
    
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        for col in range(1, 11):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _sample_fill
                cell.border = _sample_border
                if col == 1:
                    cell.value = "DEV-001"
                elif col == 2:
                    cell.value = "Firewall"
                elif col == 3:
                    cell.value = "fw-01.example.com"
            else:
                apply_style(cell, styles["input_cell"])

    ws.cell(row=start_row, column=4).value = "DMZ"
    ws.cell(row=start_row, column=5).value = "Traffic Filtering, IDS/IPS, VPN Termination"
    ws.cell(row=start_row, column=6).value = "A.8.20, A.8.22"
    ws.cell(row=start_row, column=7).value = 92
    ws.cell(row=start_row, column=8).value = "Effective"
    ws.cell(row=start_row, column=9).value = "15.01.2026"
    ws.cell(row=start_row, column=10).value = ""

    # Example Device Mappings
    row = end_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "Example Device-to-Control Mappings"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 11):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Device Type"
    ws["B" + str(row)] = "Controls Provided"
    ws["C" + str(row)] = "Control Category"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    examples = [
        ("Firewall", "Perimeter Firewall, Inter-Zone Rules, Logging", "A.8.20, A.8.22"),
        ("Router", "ACL Enforcement, Network Segmentation", "A.8.20, A.8.22"),
        ("Switch", "VLAN Segmentation, Port Security, DHCP Snooping", "A.8.22"),
        ("DNS Server", "DNS Security (DNSSEC), Query Logging", "A.8.21"),
        ("DHCP Server", "DHCP Security, Scope Management", "A.8.21"),
        ("Proxy", "Web Filtering, SSL Inspection, Malware Scanning", "A.8.21"),
        ("Load Balancer", "Load Balancer Security, SSL Termination", "A.8.21"),
        ("SIEM", "Network Monitoring, Logging", "A.8.20"),
    ]
    
    row += 1
    for device, controls, category in examples:
        ws[f"A{row}"] = device
        ws[f"B{row}"] = controls
        ws[f"C{row}"] = category
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [12, 15, 20, 18, 40, 18, 12, 15, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SHEET 5 - SERVICE CONTROL MAPPING
# ============================================================================

def create_service_control_mapping_sheet(ws, styles):
    """Create Service Control Mapping - which services implement which controls."""
    
    ws.title = "Service Control Mapping"
    
    # Title
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "SERVICE-TO-CONTROL MAPPING (FROM WB3)"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:I2")
    ws["A2"] = "Map which network services implement which security controls. Data from WB3 (Services Catalog)."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Service ID",           # A (from WB3)
        "Service Type",         # B
        "Service Name",         # C
        "Controls Provided",    # D
        "Control Category",     # E (A.8.21 primarily)
        "Security Score",       # F (from WB3)
        "Zones Served",         # G
        "Last Assessed",        # H
        "Notes",                # I
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows: 1 F2F2F2 sample + 50 FFFFCC empty = 51 rows
    start_row = 4
    end_row = start_row + SERVICE_MAPPING_ROWS - 1

    _scm_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _scm_thin = Side(style="thin")
    _scm_border = Border(left=_scm_thin, right=_scm_thin, top=_scm_thin, bottom=_scm_thin)
    scm_sample_vals = ["SCM-001", "DNS", "Core DNS Server", "DNS Security, DNSSEC, Query Filtering", "A.8.21", 92, "DMZ, Management", "15.01.2026", ""]

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = scm_sample_vals[col - 1]
                cell.fill = _scm_sample_fill
                cell.border = _scm_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])
    
    # Service Control Examples
    row = end_row + 2
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "Example Service-to-Control Mappings"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 10):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Service"
    ws["B" + str(row)] = "Controls Provided"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    examples = [
        ("DNS", "DNS Security (DNSSEC), Query Logging, Rate Limiting, Split DNS"),
        ("DHCP", "DHCP Security, DHCP Snooping, Rogue Detection"),
        ("NTP", "NTP Security, Time Synchronization, Authentication"),
        ("Proxy", "Web Filtering, SSL Inspection, Malware Scanning, Logging"),
        ("Load Balancer", "Load Balancing Security, SSL Termination, Health Checks"),
        ("RADIUS/TACACS+", "AAA Services, Strong Authentication, Command Authorisation"),
        ("SNMP", "SNMP Security (SNMPv3), Device Monitoring"),
        ("Syslog", "Centralised Logging, TLS Encryption, Retention"),
    ]
    
    row += 1
    for service, controls in examples:
        ws[f"A{row}"] = service
        ws.merge_cells(f"B{row}:I{row}")
        ws[f"B{row}"] = controls
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [12, 20, 25, 45, 18, 12, 20, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 9: SHEET 6 - CONTROL EFFECTIVENESS
# ============================================================================

def create_control_effectiveness_sheet(ws, styles, validations):
    """Create Control Effectiveness assessment - overall effectiveness per control."""
    
    ws.title = "Control Effectiveness"
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "OVERALL CONTROL EFFECTIVENESS ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "Assess effectiveness of each security control across all zones."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Control Name",             # A
        "Control Category",         # B (A.8.20/21/22)
        "Zones Covered",            # C (count)
        "Total Zones",              # D (count)
        "Coverage %",               # E - Auto
        "Effectiveness Rating",     # F - Dropdown
        "Weaknesses",               # G
        "Improvement Actions",      # H
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Pre-populate control names
    row = 4
    for category, controls in CONTROL_CATEGORIES.items():
        for control in controls:
            ws.cell(row=row, column=1, value=control)
            ws.cell(row=row, column=2, value=category)
            
            # Input cells for columns C onwards
            for col in range(3, 9):
                cell = ws.cell(row=row, column=col)
                apply_style(cell, styles["input_cell"])
            
            row += 1
    
    start_row = 4
    end_row = row - 1
    
    # Apply Data Validations
    validations["effectiveness"].add(f"F{start_row}:F{end_row}")
    ws.add_data_validation(validations["effectiveness"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Effective"'], fill=styles["covered_fill"]["fill"], font=styles["covered_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Ineffective"'], fill=styles["not_covered_fill"]["fill"], font=styles["not_covered_fill"]["font"])
    )
    
    # Column widths
    widths = [30, 30, 12, 12, 12, 18, 35, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: SHEET 7 - COVERAGE GAPS
# ============================================================================

def create_coverage_gaps_sheet(ws, styles, validations):
    """Create Gap Analysis sheet - zones/assets with insufficient controls."""
    
    ws.title = "Gap Analysis"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "SECURITY CONTROL COVERAGE GAPS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "Zones or assets with insufficient security control coverage."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",               # A - Auto
        "Zone/Asset",           # B
        "Missing Control",      # C
        "Control Category",     # D
        "Risk",                 # E
        "Severity",             # F - Dropdown
        "Remediation Plan",     # G
        "Owner",                # H
        "Target Date",          # I
        "Status",               # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1
    
    _cov_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _cov_yell = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _cov_bord = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        # Gap ID (auto-generated)
        gap_num = row_idx - start_row + 1
        is_sample = (row_idx == start_row)
        c1 = ws.cell(row=row_idx, column=1)
        if is_sample:
            c1.value = "COV-GAP-001"
            c1.fill = _cov_grey
            c1.border = _cov_bord
            c1.font = Font(color="808080", name="Calibri")
        else:
            c1.value = f'=IF(B{row_idx}<>"","COV-GAP-" & TEXT({gap_num},"000"),"")'
            c1.fill = _cov_yell
            c1.border = _cov_bord

        # Input cells
        for col in range(2, 11):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.fill = _cov_grey
                cell.border = _cov_bord
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample data for row 4
    ws.cell(row=start_row, column=2).value = "Internet DMZ"
    ws.cell(row=start_row, column=3).value = "Missing IDS/IPS Coverage"
    ws.cell(row=start_row, column=4).value = "Preventive"
    ws.cell(row=start_row, column=5).value = "No intrusion detection on DMZ — attacks undetected"
    ws.cell(row=start_row, column=6).value = "Critical"
    ws.cell(row=start_row, column=7).value = "Deploy IDS/IPS at DMZ perimeter; configure alerts to SIEM"
    ws.cell(row=start_row, column=8).value = "Security Team"
    ws.cell(row=start_row, column=9).value = "28.02.2026"
    ws.cell(row=start_row, column=10).value = "Open"

    # Apply Data Validations (exclude sample row)
    validations["gap_severity"].add(f"F{start_row + 1}:F{end_row}")
    ws.add_data_validation(validations["gap_severity"])

    validations["control_category"].add(f"D{start_row + 1}:D{end_row}")
    ws.add_data_validation(validations["control_category"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    
    # Common Gap Types
    row = end_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "Common Coverage Gap Types"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 11):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    gap_types = [
        "Missing Perimeter Firewall: Zone exposed directly to Internet without firewall (Critical)",
        "No Network Monitoring: Zone traffic not monitored by SIEM/IDS (High)",
        "Missing Segmentation: Critical zone not segregated from general network (High)",
        "No Service Security: DNS/DHCP services not hardened (Medium)",
        "Insufficient Logging: Key controls not logging events (Medium)",
        "Single Layer Defence: Only one control protecting zone (need defense-in-depth) (High)",
    ]
    
    for gap_type in gap_types:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = gap_type
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [15, 20, 30, 25, 35, 12, 40, 20, 12, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 11: SHEET 8 - DEFENSE IN DEPTH
# ============================================================================

def create_defense_in_depth_sheet(ws, styles):
    """Create Defense-in-Depth validation - multiple layers of controls."""
    
    ws.title = "Defense In Depth"
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "DEFENSE-IN-DEPTH VALIDATION"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "Verify multiple layers of controls protect critical zones (no single point of failure)."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Zone/Asset",           # A
        "Risk Level",           # B (Critical/High/Medium/Low)
        "Layer 1 Control",      # C (e.g., Perimeter Firewall)
        "Layer 2 Control",      # D (e.g., VLAN Segmentation)
        "Layer 3 Control",      # E (e.g., ACLs)
        "Additional Layers",    # F
        "Defense-in-Depth OK",  # G (Yes/No)
        "Notes",                # H
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows: 1 F2F2F2 sample + 50 FFFFCC empty = 51 rows
    start_row = 4
    end_row = start_row + DEFENSE_LAYER_COUNT - 1

    _did_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _did_thin = Side(style="thin")
    _did_border = Border(left=_did_thin, right=_did_thin, top=_did_thin, bottom=_did_thin)
    did_sample_vals = ["Internet DMZ", "Critical", "Perimeter Firewall", "VLAN Segmentation", "ACLs on Routers", "IDS/IPS Monitoring", "Yes", "Firewall + VLAN + ACL + IDS = 4 layers"]

    for row_idx in range(start_row, end_row + 1):
        is_sample = (row_idx == start_row)
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            if is_sample:
                cell.value = did_sample_vals[col - 1]
                cell.fill = _did_sample_fill
                cell.border = _did_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])
    
    # Defense-in-Depth Principles
    row = end_row + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Defense-in-Depth Principles"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 9):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    principles = [
        "• MULTIPLE LAYERS: Critical zones require 3+ independent layers of controls",
        "• NO SINGLE POINT OF FAILURE: If one control fails, others still protect",
        "• DIVERSE CONTROLS: Different types (Firewall + VLAN + ACL + Monitoring)",
        "• COMPENSATING CONTROLS: If primary control unavailable, compensating controls in place",
        "• CRITICAL ASSETS: Datacenter, management networks require strongest defense-in-depth",
    ]
    
    for principle in principles:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = principle
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Example Layers
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Example Defense-in-Depth Layers"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 9):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    examples = [
        "Layer 1: Perimeter Firewall (blocks external threats)",
        "Layer 2: Network Segmentation (isolates zones via VLANs)",
        "Layer 3: Internal Firewalls (zone-to-zone filtering)",
        "Layer 4: ACLs on Routers/Switches (additional filtering)",
        "Layer 5: Network Monitoring/IDS (detect anomalies)",
        "Layer 6: Host-based Firewalls (endpoint protection)",
    ]
    
    for example in examples:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = example
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [25, 12, 25, 25, 25, 30, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 12: SHEET 9 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1, TABLE 2, TABLE 3."""

    ws.title = "Summary Dashboard"

    _SQ = "'"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _white_bold_14 = Font(bold=True, color="FFFFFF", size=14)
    _white_bold_11 = Font(bold=True, color="FFFFFF", size=11)
    _dark_bold_10 = Font(bold=True, color="000000", size=10)
    _dark_10 = Font(bold=False, color="000000", size=10)
    _dark_9_italic = Font(bold=False, color="000000", size=9, italic=True)
    _title_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _data_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _t3total_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _t3_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    _center = Alignment(horizontal="center", vertical="center")
    _left = Alignment(horizontal="left", vertical="center")
    _wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _banner(row, text, font, fill):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = font
        c.fill = fill

    def _hdr_row_t1(row):
        for col_letter, text in [("A", "Assessment Area"), ("B", "Total"), ("C", "Yes"), ("D", "Partial"), ("E", "No"), ("F", "N-A"), ("G", "Compliance %")]:
            c = ws[f"{col_letter}{row}"]
            c.value = text
            c.font = _dark_bold_10
            c.fill = _hdr_fill
            c.border = _thin
            c.alignment = _center

    def _data_row(row, area, b, c_val, d, e, f_val, g):
        ws[f"A{row}"].value = area
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col, val in [("B", b), ("C", c_val), ("D", d), ("E", e), ("F", f_val)]:
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.font = _dark_10
            cell.border = _thin
            cell.alignment = _center
        ws[f"G{row}"].value = g
        ws[f"G{row}"].font = _dark_10
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    def _total_row_t1(row, first, last):
        ws[f"A{row}"].value = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"A{row}"].fill = _total_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col in ["B", "C", "D", "E", "F"]:
            c = ws[f"{col}{row}"]
            c.value = f"=SUM({col}{first}:{col}{last})"
            c.font = Font(bold=True, color="000000", size=10)
            c.fill = _total_fill
            c.border = _thin
            c.alignment = _center
        ws[f"G{row}"].value = f"=IF((B{row}-F{row})=0,\"0%\",ROUND(C{row}/(B{row}-F{row})*100,1)&\"%\")"
        ws[f"G{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"G{row}"].fill = _total_fill
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    # ── ROW 1: Title ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 35
    _banner(1, "NETWORK SECURITY CONTROLS COVERAGE MATRIX — SUMMARY DASHBOARD", _white_bold_14, _title_fill)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── ROW 2: Subtitle ───────────────────────────────────────────────
    ws["A2"] = "ISO/IEC 27001:2022 — Controls A.8.20 · A.8.21 · A.8.22: Network Security, Security of Network Services and Segregation of Networks"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = _left
    ws.merge_cells("A2:G2")

    # ── TABLE 1 ───────────────────────────────────────────────────────
    _banner(4, "TABLE 1: ASSESSMENT AREA COMPLIANCE", _white_bold_11, _title_fill)
    _hdr_row_t1(5)

    # Row 6: Controls Coverage Matrix (25 cols C:AA × ZONE_ROW_COUNT=51 zones, rows 4-54)
    _data_row(6,
        "Controls Coverage Matrix",
        f"=COUNTA({_SQ}Controls Coverage Matrix{_SQ}!C5:AA54)",
        f"=COUNTIF({_SQ}Controls Coverage Matrix{_SQ}!C5:AA54,\"Covered\")",
        f"=COUNTIF({_SQ}Controls Coverage Matrix{_SQ}!C5:AA54,\"Partially Covered\")",
        f"=COUNTIF({_SQ}Controls Coverage Matrix{_SQ}!C5:AA54,\"Not Covered\")",
        "=B6-(C6+D6+E6)",
        "=IF((B6-F6)=0,\"0%\",ROUND(C6/(B6-F6)*100,1)&\"%\")"
    )

    # Row 7: Zone Control Assessment (col G Effectiveness Rating, rows 4-54)
    _data_row(7,
        "Zone Control Assessment",
        f"=COUNTA({_SQ}Zone Control Assessment{_SQ}!B5:B54)",
        f"=COUNTIF({_SQ}Zone Control Assessment{_SQ}!G5:G54,\"Effective\")",
        f"=COUNTIF({_SQ}Zone Control Assessment{_SQ}!G5:G54,\"Partially Effective\")",
        f"=COUNTIF({_SQ}Zone Control Assessment{_SQ}!G5:G54,\"Ineffective\")",
        "=B7-(C7+D7+E7)",
        "=IF((B7-F7)=0,\"0%\",ROUND(C7/(B7-F7)*100,1)&\"%\")"
    )

    # Row 8: Device Control Mapping (col H Effectiveness, rows 4-54)
    _data_row(8,
        "Device Control Mapping",
        f"=COUNTA({_SQ}Device Control Mapping{_SQ}!B5:B54)",
        f"=COUNTIF({_SQ}Device Control Mapping{_SQ}!H5:H54,\"Effective\")",
        f"=COUNTIF({_SQ}Device Control Mapping{_SQ}!H5:H54,\"Partially Effective\")",
        f"=COUNTIF({_SQ}Device Control Mapping{_SQ}!H5:H54,\"Ineffective\")",
        "=B8-(C8+D8+E8)",
        "=IF((B8-F8)=0,\"0%\",ROUND(C8/(B8-F8)*100,1)&\"%\")"
    )

    # Row 9: Service Control Mapping (col F Security Score, rows 4-54 = SERVICE_MAPPING_ROWS=51)
    _data_row(9,
        "Service Control Mapping",
        f"=COUNTA({_SQ}Service Control Mapping{_SQ}!B5:B54)",
        f"=COUNTIF({_SQ}Service Control Mapping{_SQ}!F5:F54,\">=80\")",
        f"=COUNTIFS({_SQ}Service Control Mapping{_SQ}!F5:F54,\">=50\",{_SQ}Service Control Mapping{_SQ}!F5:F54,\"<80\")",
        f"=COUNTIFS({_SQ}Service Control Mapping{_SQ}!F5:F54,\">=0\",{_SQ}Service Control Mapping{_SQ}!F5:F54,\"<50\")",
        "=B9-(C9+D9+E9)",
        "=IF((B9-F9)=0,\"0%\",ROUND(C9/(B9-F9)*100,1)&\"%\")"
    )

    # Row 10: Control Effectiveness (col F Effectiveness Rating, rows 4-24 = 21 controls)
    _data_row(10,
        "Control Effectiveness",
        f"=COUNTA({_SQ}Control Effectiveness{_SQ}!B5:B24)",
        f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Effective\")",
        f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Partially Effective\")",
        f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Ineffective\")",
        "=B10-(C10+D10+E10)",
        "=IF((B10-F10)=0,\"0%\",ROUND(C10/(B10-F10)*100,1)&\"%\")"
    )

    # Row 11: TOTAL
    _total_row_t1(11, 6, 10)

    # ── TABLE 2 ───────────────────────────────────────────────────────
    _banner(13, "TABLE 2: KEY METRICS", _white_bold_11, _title_fill)

    ws["A14"].value = "Metric"
    ws["A14"].font = _dark_bold_10
    ws["A14"].fill = _hdr_fill
    ws["A14"].border = _thin
    ws["A14"].alignment = _left
    ws["B14"].value = "Value"
    ws["B14"].font = _dark_bold_10
    ws["B14"].fill = _hdr_fill
    ws["B14"].border = _thin
    ws["B14"].alignment = _center
    ws.merge_cells("C14:G14")
    ws["C14"].value = "What This Shows"
    ws["C14"].font = _dark_bold_10
    ws["C14"].fill = _hdr_fill
    ws["C14"].border = _thin
    ws["C14"].alignment = _left

    t2_metrics = [
        (15, "Uncovered Control-Zone Combinations",
         f"=COUNTIF({_SQ}Controls Coverage Matrix{_SQ}!C5:AA54,\"Not Covered\")",
         "Zones without required security controls (defence-in-depth gaps)"),
        (16, "Partially Covered Combinations",
         f"=COUNTIF({_SQ}Controls Coverage Matrix{_SQ}!C5:AA54,\"Partially Covered\")",
         "Zones with incomplete security control coverage"),
        (17, "Ineffective Controls",
         f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Ineffective\")",
         "Controls assessed as ineffective (non-functional or bypassed)"),
        (18, "Partially Effective Controls",
         f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Partially Effective\")",
         "Controls operating at reduced effectiveness"),
        (19, "Not Assessed Controls",
         f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Not Assessed\")",
         "Controls without effectiveness rating (assessment required)"),
        (20, "Critical Gap Analysis",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!F5:F43,\"Critical\")",
         "Critical-severity control coverage gaps (immediate deployment required)"),
        (21, "High Severity Gap Analysis",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!F5:F43,\"High\")",
         "High-severity control gaps requiring prioritised remediation"),
        (22, "Overall Compliance Rate",
         "=G11",
         "Overall controls coverage compliance percentage (TABLE 1 total)"),
    ]

    for row, metric, formula, description in t2_metrics:
        ws[f"A{row}"].value = metric
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = description
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).border = _thin

    # ── TABLE 3 ───────────────────────────────────────────────────────
    _banner(24, "TABLE 3: CRITICAL FINDINGS", _white_bold_11, _t3_fill)

    ws["A25"].value = "Critical Finding Type"
    ws["A25"].font = _dark_bold_10
    ws["A25"].fill = _hdr_fill
    ws["A25"].border = _thin
    ws["A25"].alignment = _left
    ws["B25"].value = "Count"
    ws["B25"].font = _dark_bold_10
    ws["B25"].fill = _hdr_fill
    ws["B25"].border = _thin
    ws["B25"].alignment = _center
    ws.merge_cells("C25:G25")
    ws["C25"].value = "Filter Instructions"
    ws["C25"].font = _dark_bold_10
    ws["C25"].fill = _hdr_fill
    ws["C25"].border = _thin
    ws["C25"].alignment = _left

    t3_findings = [
        (26, "Uncovered Control-Zone Combinations",
         f"=COUNTIF({_SQ}Controls Coverage Matrix{_SQ}!C5:AA54,\"Not Covered\")",
         "Filter Controls Coverage Matrix: Any control column = \"Not Covered\""),
        (27, "Ineffective Controls",
         f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Ineffective\")",
         "Filter Control Effectiveness: Effectiveness Rating = \"Ineffective\""),
        (28, "Critical Gap Analysis",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!F5:F43,\"Critical\")",
         "Filter Gap Analysis: Severity = \"Critical\""),
        (29, "High Severity Gap Analysis",
         f"=COUNTIF({_SQ}Gap Analysis{_SQ}!F5:F43,\"High\")",
         "Filter Gap Analysis: Severity = \"High\""),
        (30, "Not Assessed Controls",
         f"=COUNTIF({_SQ}Control Effectiveness{_SQ}!F5:F24,\"Not Assessed\")",
         "Filter Control Effectiveness: Effectiveness Rating = \"Not Assessed\""),
    ]

    for row, finding_type, formula, instructions in t3_findings:
        ws[f"A{row}"].value = finding_type
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].fill = _data_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].fill = _data_fill
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = instructions
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].fill = _data_fill
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).fill = _data_fill
            ws.cell(row=row, column=_c).border = _thin

    # TOTAL row
    ws["A31"].value = "TOTAL"
    ws["A31"].font = Font(bold=True, color="000000", size=10)
    ws["A31"].border = _thin
    ws["A31"].alignment = _left
    ws["B31"].value = "=SUM(B26:B30)"
    ws["B31"].font = Font(bold=True, color="000000", size=10)
    ws["B31"].fill = _t3total_fill
    ws["B31"].border = _thin
    ws["B31"].alignment = _center
    ws.merge_cells("C31:G31")
    ws["C31"].value = "Total critical findings requiring immediate investigation"
    ws["C31"].font = Font(italic=True, size=9, color="000000")
    ws["C31"].alignment = _left

    # ── Column widths ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 13: SHEET 10 - EXECUTIVE SUMMARY
# ============================================================================

def create_executive_summary_sheet(ws, styles):
    """Create Executive Summary for management."""
    
    ws.title = "Executive Summary"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "NETWORK SECURITY CONTROLS - EXECUTIVE SUMMARY"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Summary sections
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Assessment Overview"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    overview = [
        "This workbook (WB5) integrates findings from:",
        "  • WB1: Network Infrastructure Inventory",
        "  • WB2: Device Security Assessment",
        "  • WB3: Network Services Catalog",
        "  • WB4: Network Segmentation Matrix",
        "",
        "Purpose: Map security controls coverage across all network zones and assets.",
        "Controls Assessed: ISO 27001:2022 A.8.20 (Network Security), A.8.21 (Network Services), A.8.22 (Network Segmentation)",
    ]
    
    for line in overview:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = line
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Key Findings
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Key Findings"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    findings = [
        "1. COVERAGE: [Insert average zone coverage %] average control coverage across all zones",
        "2. GAPS: [Insert count] critical coverage gaps identified (see Gap Analysis sheet)",
        "3. EFFECTIVENESS: [Insert %] of controls assessed as 'Effective'",
        "4. DEFENSE-IN-DEPTH: [Insert count] critical zones with insufficient layered controls",
        "5. COMPLIANCE: Overall compliance with A.8.20/21/22 is [Insert status]",
    ]
    
    for finding in findings:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = finding
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Recommendations
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Priority Recommendations"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    recommendations = [
        "1. IMMEDIATE: Remediate all Critical coverage gaps (see Gap Analysis sheet)",
        "2. HIGH PRIORITY: Implement defense-in-depth for all critical zones (Datacenter, Management)",
        "3. MEDIUM PRIORITY: Improve control effectiveness for zones with 'Partially Effective' rating",
        "4. ONGOING: Maintain Controls Coverage Matrix quarterly, reassess after major changes",
    ]
    
    for rec in recommendations:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = rec
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Next Steps
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Next Steps"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 7):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    next_steps = [
        "1. Review Controls Coverage Matrix with IT Security team",
        "2. Prioritize remediation of Critical gaps",
        "3. Assign owners for all gaps in Gap Analysis sheet",
        "4. Schedule follow-up assessment in [3 months / 6 months]",
        "5. Update controls matrix after each major network change",
    ]
    
    for step in next_steps:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = step
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20


# ============================================================================
# SECTION 14: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb, styles):
    """Create Evidence Register sheet — golden standard."""
    ws = wb["Evidence Register"]
    ws.sheet_view.showGridLines = False
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (italic, NOT blue banner) ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers (003366, white text) ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Controls Coverage Assessment",
        3: "Configuration file",
        4: "Network control coverage matrix and zone control assessment documentation",
        5: "/evidence/A.8.20-22/",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"
def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws = wb["Approval Sign-Off"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE(\'Summary Dashboard\'!G6:G10),\"\")")  ,
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    styles = _STYLES
    validations = create_data_validations()

    create_instructions_sheet(wb.create_sheet())
    create_controls_coverage_matrix_sheet(wb.create_sheet(), styles, validations)
    create_zone_control_assessment_sheet(wb.create_sheet(), styles, validations)
    create_device_control_mapping_sheet(wb.create_sheet(), styles)
    create_service_control_mapping_sheet(wb.create_sheet(), styles)
    create_control_effectiveness_sheet(wb.create_sheet(), styles, validations)
    create_coverage_gaps_sheet(wb.create_sheet(), styles, validations)
    create_defense_in_depth_sheet(wb.create_sheet(), styles)
    create_executive_summary_sheet(wb.create_sheet(), styles)

    wb.create_sheet("Evidence Register")
    create_evidence_register(wb, styles)

    create_summary_dashboard_sheet(wb.create_sheet(), styles)

    wb.create_sheet("Approval Sign-Off")
    create_approval_sheet(wb)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path.name}")

def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"ERROR saving workbook: {e}")
        sys.exit(1)
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\nCore Assessment Sheets:")
    logger.info(f"  • Controls Coverage Matrix ({ZONE_ROW_COUNT} zones × {CONTROL_COLUMNS} controls)")
    logger.info(f"  • Zone Control Assessment ({ZONE_ROW_COUNT} zones)")
    logger.info(f"  • Device Control Mapping ({DEVICE_MAPPING_ROWS} devices)")
    logger.info(f"  • Service Control Mapping ({SERVICE_MAPPING_ROWS} services)")
    logger.info(f"  • Control Effectiveness ({CONTROL_COLUMNS} controls)")
    logger.info("\n Analysis & Reporting:")
    logger.info(f"  • Gap Analysis ({GAP_ROW_COUNT} gaps)")
    logger.info(f"  • Defense In Depth ({DEFENSE_LAYER_COUNT} layers)")
    logger.info("  • Compliance Dashboard (aggregated metrics)")
    logger.info("  • Executive Summary (management view)")
    logger.info("  • Evidence Register (100 evidence rows)")
    logger.info("  • Approval Sign-Off (3 approvers)")
    logger.info("\n" + "=" * 80)
    logger.info("NEXT STEPS:")
    logger.info("  1. Complete Controls Coverage Matrix (zone × control mapping)")
    logger.info("  2. Map devices to controls (from WB1 + WB2)")
    logger.info("  3. Map services to controls (from WB3)")
    logger.info("  4. Assess control effectiveness per zone")
    logger.info("  5. Identify coverage gaps")
    logger.info("  6. Validate defense-in-depth for critical zones")
    logger.info("  7. Review Compliance Dashboard and Executive Summary")
    logger.info("\n" + "=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
