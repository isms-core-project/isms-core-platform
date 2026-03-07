#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-22.S4 - Network Segmentation Matrix Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 4 of 6: Network Segregation and Security Zones

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific network segmentation architecture, security zones,
and assessment requirements.

Key customization areas:
1. Security zone definitions (match your actual segmentation architecture)
2. Trust boundary requirements (based on your risk assessment)
3. Inter-zone traffic policies (aligned with your business needs)
4. Segmentation technologies (specific to your infrastructure)
5. Effectiveness testing criteria (adapted to your security standards)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
network segmentation architecture, security zones, trust boundaries, and
inter-zone traffic controls across the organisation.

**Purpose:**
Enables comprehensive assessment of network segmentation architecture, security
zones, trust boundaries, and inter-zone traffic controls, supporting evidence-based
validation of network segregation effectiveness against ISO 27001:2022 Control A.8.22.

**Assessment Scope:**
- Security zones inventory and definition
- Zone purpose and classification (DMZ, internal, management, guest)
- Networks/subnets per zone (VLAN assignments, IP ranges)
- Trust relationships between zones
- Inter-zone traffic policies (firewall rules, ACLs)
- Segmentation technology (VLAN, VRF, physical, virtualization)
- Effectiveness testing results
- Flat network identification and remediation requirements

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and segmentation standards
2. Security Zones Inventory - Security zones inventory and definitions
3. Segmentation Matrix - Zone-to-zone trust relationships and traffic matrix
4. VLAN Inventory - VLAN inventory and VLAN-to-zone mapping
5. Firewall Rules Assessment - Firewall rules inventory and zone protection
6. ACL Assessment - Router/switch ACL assessment
7. Segmentation Testing - Segmentation effectiveness test results
8. Gap Analysis - Segmentation gaps and remediation tracking
9. Summary Dashboard - Metrics, charts, and compliance targets
10. Remediation Roadmap - Prioritised remediation actions
11. Evidence Register - Audit evidence tracking and documentation
12. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with zone classification dropdown lists
- Conditional formatting for segmentation effectiveness status
- Automated zone-to-zone traffic matrix generation
- Protected formulas with unprotected input cells
- Segmentation technology tracking (VLAN, VRF, physical)
- Flat network identification and risk assessment
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with firewall management systems

**Integration:**
This assessment works with WB1 (Infrastructure Inventory) to map devices to
security zones, integrates with WB2 (Device Security) for zone protection controls,
and feeds into WB5 (Controls Coverage) and the Network Security Compliance
Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_4_segmentation_matrix.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_4_segmentation_matrix.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_4_segmentation_matrix.py --date 20250124
    
    # Generate with custom organisation name
    python3 generate_a820_4_segmentation_matrix.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organisation name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_4_Segmentation_Matrix_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize security zone definitions to match your architecture
    2. Document all security zones and their purpose/classification
    3. Map networks/subnets to security zones
    4. Define inter-zone trust relationships and traffic policies
    5. Inventory firewall rules protecting zone boundaries
    6. Document segmentation technology (VLAN, VRF, physical separation)
    7. Conduct segmentation effectiveness testing
    8. Identify flat networks and assess remediation requirements
    9. Collect and link audit evidence (firewall configs, test results)
    10. Obtain stakeholder approvals
    11. Feed results into WB5 and Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    4 of 6 (Network Segregation and Security Zones)
Primary Control:      A.8.22 (Segregation of Networks)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalisation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-22 specification
    - Supports comprehensive network segmentation assessment
    - Integrated with A.8.20-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Network Segmentation Standards:**
Network segmentation architecture should follow security best practices:
- Defense in depth with multiple security zones
- Least privilege for inter-zone traffic
- DMZ for internet-facing services
- Separate management network for device administration
- Guest network isolation
- Microsegmentation for critical assets

Review and update segmentation architecture based on changing risk landscape.

**Technology Diversity:**
This assessment framework is technology-agnostic and must work across:
- Traditional networks (VLANs, physical separation, router ACLs)
- Software-Defined Networks (SDN policies, virtual networks)
- Cloud environments (AWS VPCs, Azure VNets, GCP VPCs, security groups)
- Hybrid architectures (on-premise zones + cloud security groups)
- Microsegmentation (VMware NSX, Cisco ACI, Illumio)

Customize assessment criteria to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Documented security zone architecture with clear purpose/classification
- Inter-zone traffic policies (firewall rules, ACLs)
- Evidence of segmentation effectiveness (test results)
- Flat network identification and risk acceptance or remediation plans
- Trust boundary documentation and controls

**Data Protection:**
Assessment workbooks contain sensitive network architecture information including:
- Complete network segmentation architecture
- Security zone boundaries and trust relationships
- Firewall rules and traffic policies
- Identified flat networks and segmentation gaps

Handle in accordance with your organisation's data classification policies.
Restrict access to authorised network architects and security personnel only.

**Maintenance:**
Review and update segmentation assessment:
- Monthly: After significant network changes (new zones, zone modifications)
- Quarterly: Routine reassessment of inter-zone traffic policies
- Annually: Complete segmentation architecture review
- Ad-hoc: After security incidents or penetration test findings

**Quality Assurance:**
Validate segmentation effectiveness by:
- Testing inter-zone traffic filtering (authorised vs. unauthorised traffic)
- VLAN hopping testing
- Firewall rule validation and cleanup
- Network topology verification against segmentation architecture
- Penetration testing across zone boundaries
- Flat network scanning and identification

**Regulatory Alignment:**
Ensure network segmentation aligns with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 network segmentation requirements (cardholder data)
- Healthcare: HIPAA network segmentation for ePHI
- Finance: Regional banking network segmentation requirements
- Government: Jurisdiction-specific segmentation mandates

Customize assessment criteria to include regulatory-specific requirements.

**Integration Points:**
This assessment integrates with other ISO 27001 controls:
- A.8.20 (Network Security): Devices protecting zone boundaries
- A.8.21 (Network Services): Services per security zone
- A.8.15 (Logging): Inter-zone traffic logging
- A.8.16 (Monitoring): Zone boundary monitoring

**Common Pitfalls to Avoid:**
1. **Flat Networks**: Everything in one VLAN/subnet (no segmentation)
2. **Overly Permissive Rules**: "Allow Any Any" between zones
3. **No Testing**: Assuming segmentation works without testing
4. **Documentation Drift**: Topology diagrams don't match reality
5. **Missing DMZ**: Internet-facing services in internal network
6. **No Management Network**: Admin traffic mixed with production
7. **VLAN Hopping**: Not preventing VLAN hopping attacks
8. **Cloud Confusion**: Assuming cloud providers handle segmentation

**Segmentation Effectiveness Testing:**

**Zone Boundary Testing:**
- Verify firewall rules block unauthorised traffic between zones
- Test from each zone to every other zone (zone-to-zone matrix)
- Validate both allowed and denied traffic
- Document test methodology and results

**VLAN Security Testing:**
- VLAN hopping attempts (double tagging, switch spoofing)
- Trunk port security verification
- Native VLAN security
- Private VLAN (PVLAN) effectiveness if used

**Microsegmentation Testing (if applicable):**
- Host-to-host traffic filtering
- Application-level segmentation
- East-west traffic controls
- Policy enforcement verification

**Flat Network Detection:**
- Identify networks without segmentation
- Assess risk of flat network architectures
- Develop remediation plans or risk acceptance

**Common Security Zone Types:**

**Internet DMZ:**
- Purpose: Internet-facing services (web servers, mail relays)
- Trust Level: Untrusted
- Controls: Strict ingress/egress filtering, IDS/IPS, web application firewall

**Internal Network:**
- Purpose: Corporate workstations, internal servers
- Trust Level: Trusted
- Controls: Access control, monitoring, endpoint protection

**Management Network:**
- Purpose: Device administration (SSH, RDP, management interfaces)
- Trust Level: Restricted
- Controls: Jump hosts, MFA, privileged access management

**Server/Data Center Network:**
- Purpose: Production servers, databases, applications
- Trust Level: Trusted (with microsegmentation)
- Controls: Application-level firewalls, database firewalls

**Guest Network:**
- Purpose: Visitor/guest wireless access
- Trust Level: Untrusted
- Controls: Complete isolation from internal networks, captive portal

**Voice/Video Network:**
- Purpose: VoIP phones, video conferencing
- Trust Level: Segmented
- Controls: QoS, separate VLAN, restricted access

**IoT/OT Network:**
- Purpose: IoT devices, operational technology
- Trust Level: Segmented
- Controls: Strict isolation, dedicated monitoring, anomaly detection

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "Network Segmentation Matrix & Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.20-22.S4"
CONTROL_ID   = "A.8.20-22"
CONTROL_NAME = "Network Security"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

DASH = "  -  "
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Assessment constants
ZONE_ROW_COUNT = 50             # Security zones
VLAN_ROW_COUNT = 100            # VLAN inventory
FIREWALL_RULE_COUNT = 100       # Firewall rules
ACL_COUNT = 50                  # ACLs
TEST_ROW_COUNT = 30             # Segmentation tests
GAP_ROW_COUNT = 50              # Gap tracking
REMEDIATION_ROW_COUNT = 30      # Remediation actions

# Common security zones (examples)
COMMON_ZONES = [
    "Internet/DMZ",
    "Internal",
    "Management",
    "Datacenter",
    "Guest",
    "Branch",
    "Cloud",
]
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "allow_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="006100"),
        },
        "deny_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="9C0006"),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "low_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "pass_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "fail_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "info_box": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Trust Level validation
    validations["trust_level"] = DataValidation(
        type="list",
        formula1='"Untrusted,Low Trust,Medium Trust,High Trust,Trusted"',
        allow_blank=False,
    )
    
    # Traffic Action validation
    validations["traffic_action"] = DataValidation(
        type="list",
        formula1='"Allow,Deny,Inspect"',
        allow_blank=False,
    )
    
    # Segmentation Type validation
    validations["segmentation_type"] = DataValidation(
        type="list",
        formula1='"VLAN,Physical,VRF,Virtual Network,Cloud Security Group"',
        allow_blank=True,
    )
    
    # Firewall Action validation
    validations["firewall_action"] = DataValidation(
        type="list",
        formula1='"Allow,Deny,Drop,Reject"',
        allow_blank=False,
    )
    
    # Rule Review Status validation
    validations["review_status"] = DataValidation(
        type="list",
        formula1='"Current,Outdated,Unused,Requires Review"',
        allow_blank=True,
    )
    
    # Test Result validation
    validations["test_result"] = DataValidation(
        type="list",
        formula1='"Pass,Fail,Partial,Not Tested"',
        allow_blank=False,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False,
    )
    
    # Gap Status validation
    validations["gap_status"] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Accepted Risk"',
        allow_blank=False,
    )
    
    return validations


# ============================================================================
# SECTION 4: FINALIZE VALIDATIONS
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 5: SHEET 1 - INSTRUCTIONS & ASSESSMENT GUIDE
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Review and document the network segmentation architecture.', '2. Map zone-to-zone communication paths in the Segmentation Matrix.', '3. Complete the Zone Security Assessment for each network zone.', '4. Document firewall rules and ACLs governing inter-zone traffic.', '5. Assess micro-segmentation capabilities and deployment status.', '6. Review compliance against segmentation policy requirements.', '7. Identify and document gaps with remediation plans.', '8. Maintain the Evidence Register with all supporting documentation.', '9. Obtain final approval and sign-off from Network Engineering, ISO, and CISO.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_security_zones_sheet(ws, styles, validations):
    """Create Security Zones Inventory sheet."""
    
    ws.title = "Security Zones Inventory"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = f"SECURITY ZONES INVENTORY - GENERATED {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Define all security zones in the network. This is the foundation for the segmentation matrix."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Zone ID",              # A - Auto
        "Zone Name",            # B
        "Purpose",              # C
        "Trust Level",          # D - Dropdown
        "Networks/Subnets",     # E (IP ranges)
        "VLANs",                # F (VLAN IDs)
        "Segmentation Type",    # G - Dropdown
        "Devices Count",        # H
        "Critical Assets",      # I (Y/N)
        "Internet Access",      # J (Y/N)
        "Owner",                # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + ZONE_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Zone ID (auto-generated)
        zone_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","ZONE-" & TEXT({zone_num},"00"),"")')
        
        # Input cells
        for col in range(2, 13):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["trust_level"].add(f"D{start_row}:D{end_row}")
    ws.add_data_validation(validations["trust_level"])
    
    validations["segmentation_type"].add(f"G{start_row}:G{end_row}")
    ws.add_data_validation(validations["segmentation_type"])
    
    # Column widths
    widths = [12, 20, 35, 15, 25, 15, 20, 12, 12, 12, 20, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: SHEET 3 - SEGMENTATION MATRIX
# ============================================================================

def create_segmentation_matrix_sheet(ws, styles, validations):
    """Create Segmentation Matrix - zone-to-zone traffic rules."""
    
    ws.title = "Segmentation Matrix"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "NETWORK SEGMENTATION MATRIX - ZONE-TO-ZONE TRAFFIC RULES"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "Define allowed/denied traffic between security zones. This is the CORE segmentation assessment."
    apply_style(ws["A2"], styles["info_box"])
    
    # Matrix explanation
    ws.merge_cells("A3:K3")
    ws["A3"] = "FORMAT: Each row defines traffic flow from Source Zone → Destination Zone. Action: Allow/Deny/Inspect."
    apply_style(ws["A3"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Rule ID",              # A - Auto
        "Source Zone",          # B
        "Destination Zone",     # C
        "Traffic Action",       # D - Dropdown (Allow/Deny/Inspect)
        "Protocols/Ports",      # E (TCP/80, UDP/53, etc.)
        "Justification",        # F
        "Firewall Enforced",    # G (Y/N)
        "ACL Enforced",         # H (Y/N)
        "Monitoring",           # I (Y/N)
        "Last Reviewed",        # J
        "Notes",                # K
    ]
    
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows (50 rows for matrix entries)
    start_row = 5
    end_row = start_row + 49
    
    for row_idx in range(start_row, end_row + 1):
        # Rule ID (auto-generated)
        rule_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SEG-" & TEXT({rule_num},"000"),"")')
        
        # Input cells
        for col in range(2, 12):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["traffic_action"].add(f"D{start_row}:D{end_row}")
    ws.add_data_validation(validations["traffic_action"])
    
    # Conditional Formatting for Traffic Action
    ws.conditional_formatting.add(
        f"D{start_row}:D{end_row}",
        CellIsRule(operator="equal", formula=['"Allow"'], fill=styles["allow_fill"]["fill"], font=styles["allow_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"D{start_row}:D{end_row}",
        CellIsRule(operator="equal", formula=['"Deny"'], fill=styles["deny_fill"]["fill"], font=styles["deny_fill"]["font"])
    )
    
    # Example Matrix Entries
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "Common Segmentation Matrix Examples"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 12):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Source"
    ws["B" + str(row)] = "Destination"
    ws["C" + str(row)] = "Action"
    ws["D" + str(row)] = "Justification"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    examples = [
        ("Internet/DMZ", "Internal", "Deny", "External users cannot access internal network"),
        ("Internal", "Internet/DMZ", "Allow", "Internal users can access public services"),
        ("Internal", "Datacenter", "Allow", "Users need access to application servers"),
        ("Guest", "Internal", "Deny", "Guest Wi-Fi isolated from corporate network"),
        ("Guest", "Internet", "Allow", "Guest users can access internet only"),
        ("Management", "All Zones", "Allow", "Management needs access to all devices"),
        ("All Zones", "Management", "Deny", "Management zone is protected"),
    ]
    
    row += 1
    for src, dst, action, justification in examples:
        ws[f"A{row}"] = src
        ws[f"B{row}"] = dst
        ws[f"C{row}"] = action
        ws[f"D{row}"] = justification
        
        for col in ["A", "B", "D"]:
            apply_style(ws[f"{col}{row}"], styles["info_box"])
        
        if action == "Allow":
            apply_style(ws[f"C{row}"], styles["allow_fill"])
        elif action == "Deny":
            apply_style(ws[f"C{row}"], styles["deny_fill"])
        
        row += 1
    
    # Column widths
    widths = [12, 20, 20, 12, 25, 40, 15, 15, 12, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: SHEET 4 - VLAN INVENTORY
# ============================================================================

def create_vlan_inventory_sheet(ws, styles):
    """Create VLAN Inventory sheet."""
    
    ws.title = "VLAN Inventory"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "VLAN INVENTORY & DESIGN ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all VLANs. Verify VLAN numbering, naming standards, and zone assignment."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "VLAN ID",              # A
        "VLAN Name",            # B
        "Security Zone",        # C (link to Security Zones Inventory)
        "IP Subnet",            # D
        "Purpose",              # E
        "Devices/Ports",        # F
        "Trunk Allowed",        # G (Yes/No)
        "Native VLAN",          # H (Yes/No)
        "Last Reviewed",        # I
        "Notes",                # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Row 4: F2F2F2 sample row
    _sf = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _ts = Side(style="thin")
    _bs = Border(left=_ts, right=_ts, top=_ts, bottom=_ts)
    sample_vlan = ["10", "VLAN-10-USERS-LAN", "Internal", "10.10.0.0/24", "End-user workstations", "sw-access-01 Gi0/1-24", "No", "No", "", ""]
    for col_idx, val in enumerate(sample_vlan, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = _sf
        cell.border = _bs
        cell.font = Font(color="808080", name="Calibri")

    # Data rows 5-54: 50 empty FFFFCC rows (+ sample = 51 standard)
    start_row = 5
    end_row = 54

    for row_idx in range(start_row, end_row + 1):
        for col in range(1, 11):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # VLAN Design Best Practices
    row = end_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "VLAN Design Best Practices (per IMP-S5)"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 11):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    best_practices = [
        "• VLAN NUMBERING: Consistent scheme (e.g., 10-99 for users, 100-199 for servers, 200-299 for management)",
        "• VLAN NAMING: Descriptive names (e.g., VLAN-10-USERS-FLOOR1, VLAN-100-SERVERS-WEB)",
        "• NATIVE VLAN: Change from default VLAN 1, use dedicated unused VLAN",
        "• VLAN PRUNING: Only allow necessary VLANs on trunks",
        "• ONE VLAN PER ZONE: Each security zone maps to dedicated VLANs",
        "• DOCUMENTATION: Maintain VLAN database with assignments and purpose",
    ]
    
    for practice in best_practices:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = practice
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [10, 25, 20, 20, 35, 20, 12, 12, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 9: SHEET 5 - FIREWALL RULES ASSESSMENT
# ============================================================================

def create_firewall_rules_sheet(ws, styles, validations):
    """Create Firewall Rules Assessment sheet."""
    
    ws.title = "Firewall Rules Assessment"
    
    # Title
    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "FIREWALL RULES ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "Review firewall rules that enforce zone segmentation. Focus on inter-zone traffic controls."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Rule ID",              # A - Auto
        "Firewall",             # B (Device name)
        "Rule Number",          # C
        "Source Zone",          # D
        "Destination Zone",     # E
        "Source IP/Network",    # F
        "Destination IP/Network", # G
        "Service/Port",         # H
        "Action",               # I - Dropdown
        "Review Status",        # J - Dropdown
        "Last Modified",        # K
        "Reviewed By",          # L
        "Notes",                # M
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + FIREWALL_RULE_COUNT - 1
    
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        # Rule ID
        if row_idx == start_row:
            ws.cell(row=row_idx, column=1).value = "FW-RULE-001"
        else:
            rule_num = row_idx - start_row + 1
            ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","FW-RULE-" & TEXT({rule_num},"000"),"")')

        # Input cells
        for col in range(2, 14):
            cell = ws.cell(row=row_idx, column=col)
            if row_idx == start_row:
                cell.fill = _sample_fill
                cell.border = _sample_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample data for row 4
    ws.cell(row=start_row, column=1).value = "FW-RULE-001"
    ws.cell(row=start_row, column=2).value = "core-fw-01"
    ws.cell(row=start_row, column=3).value = "1"
    ws.cell(row=start_row, column=4).value = "Internet"
    ws.cell(row=start_row, column=5).value = "DMZ"
    ws.cell(row=start_row, column=6).value = "Any"
    ws.cell(row=start_row, column=7).value = "10.10.1.0/24"
    ws.cell(row=start_row, column=8).value = "TCP/80, TCP/443"
    ws.cell(row=start_row, column=9).value = "Allow"
    ws.cell(row=start_row, column=10).value = "Current"
    ws.cell(row=start_row, column=11).value = "01.01.2026"
    ws.cell(row=start_row, column=12).value = "Network Security Team"
    ws.cell(row=start_row, column=13).value = "Permit inbound HTTP/HTTPS to DMZ web servers"

    # Apply Data Validations
    validations["firewall_action"].add(f"I{start_row+1}:I{end_row}")
    ws.add_data_validation(validations["firewall_action"])

    validations["review_status"].add(f"J{start_row+1}:J{end_row}")
    ws.add_data_validation(validations["review_status"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"I{start_row}:I{end_row}",
        CellIsRule(operator="equal", formula=['"Allow"'], fill=styles["allow_fill"]["fill"], font=styles["allow_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"I{start_row}:I{end_row}",
        CellIsRule(operator="equal", formula=['"Deny"'], fill=styles["deny_fill"]["fill"], font=styles["deny_fill"]["font"])
    )
    
    # Firewall Rule Review Guidelines
    row = end_row + 2
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "Firewall Rule Review Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 14):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    guidelines = [
        "• QUARTERLY REVIEW: Review all firewall rules at least quarterly",
        "• REMOVE UNUSED: Identify and remove unused rules (hit count = 0)",
        "• SPECIFIC > ANY: Avoid 'any any any' rules; be specific about source/destination/service",
        "• LOGGING: Enable logging for inter-zone traffic (especially deny rules)",
        "• RULE ORDER: Most specific rules first, default deny last",
        "• DOCUMENTATION: Every rule must have business justification",
    ]
    
    for guideline in guidelines:
        ws.merge_cells(f"A{row}:M{row}")
        ws[f"A{row}"] = guideline
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [15, 20, 12, 18, 18, 20, 20, 18, 12, 15, 12, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: SHEET 6 - ACL ASSESSMENT
# ============================================================================

def create_acl_assessment_sheet(ws, styles, validations):
    """Create ACL (Access Control List) Assessment sheet."""
    
    ws.title = "ACL Assessment"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "ACL (ACCESS CONTROL LIST) ASSESSMENT"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Review router/switch ACLs that enforce segmentation. Complement firewall rules."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "ACL ID",               # A - Auto
        "Device",               # B (Router/Switch)
        "ACL Name/Number",      # C
        "Interface",            # D
        "Direction",            # E (Inbound/Outbound)
        "Purpose",              # F
        "Zones Protected",      # G
        "Review Status",        # H - Dropdown
        "Effectiveness",        # I (Tested Y/N)
        "Last Modified",        # J
        "Reviewed By",          # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + ACL_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # ACL ID (auto-generated)
        acl_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","ACL-" & TEXT({acl_num},"00"),"")')
        
        # Input cells
        for col in range(2, 13):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["review_status"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["review_status"])
    
    # ACL Best Practices
    row = end_row + 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "ACL Best Practices (per IMP-S5)"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 13):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    best_practices = [
        "• ACL PLACEMENT: Apply inbound on router interfaces (filter before routing)",
        "• SPECIFIC FIRST: Most specific ACL entries first, general/deny last",
        "• EXPLICIT DENY: Include explicit deny at end (implicit deny doesn't log)",
        "• NAMING: Use descriptive names for extended ACLs (e.g., ACL-BLOCK-GUEST-TO-INTERNAL)",
        "• TESTING: Test ACLs with traffic generators/packet captures before production",
        "• DOCUMENTATION: Maintain ACL documentation with business justification per entry",
    ]
    
    for practice in best_practices:
        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = practice
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [12, 20, 20, 15, 12, 35, 20, 15, 12, 12, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 11: SHEET 7 - SEGMENTATION TESTING
# ============================================================================

def create_segmentation_testing_sheet(ws, styles, validations):
    """Create Segmentation Effectiveness Testing sheet."""
    
    ws.title = "Segmentation Testing"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "SEGMENTATION EFFECTIVENESS TESTING"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "Test segmentation effectiveness. Verify traffic is blocked when expected (no bypass routes)."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Test ID",              # A - Auto
        "Test Name",            # B
        "Source Zone",          # C
        "Destination Zone",     # D
        "Expected Result",      # E (Allow/Deny)
        "Actual Result",        # F (Allow/Deny)
        "Test Result",          # G - Dropdown (Pass/Fail)
        "Test Method",          # H (ping, traceroute, nmap, etc.)
        "Test Date",            # I
        "Tested By",            # J
        "Notes",                # K
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + TEST_ROW_COUNT - 1
    
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        # Test ID
        if row_idx == start_row:
            ws.cell(row=row_idx, column=1).value = "SEG-TEST-01"
        else:
            test_num = row_idx - start_row + 1
            ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SEG-TEST-" & TEXT({test_num},"00"),"")')

        # Input cells
        for col in range(2, 12):
            cell = ws.cell(row=row_idx, column=col)
            if row_idx == start_row:
                cell.fill = _sample_fill
                cell.border = _sample_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample data for row 4
    ws.cell(row=start_row, column=1).value = "SEG-TEST-01"
    ws.cell(row=start_row, column=2).value = "Internet DMZ to Corporate LAN"
    ws.cell(row=start_row, column=3).value = "Internet DMZ"
    ws.cell(row=start_row, column=4).value = "Corporate LAN"
    ws.cell(row=start_row, column=5).value = "Deny"
    ws.cell(row=start_row, column=6).value = "Deny"
    ws.cell(row=start_row, column=7).value = "Pass"
    ws.cell(row=start_row, column=8).value = "Port Scan (nmap)"
    ws.cell(row=start_row, column=9).value = "15.01.2026"
    ws.cell(row=start_row, column=10).value = "Network Security Team"
    ws.cell(row=start_row, column=11).value = "Firewall confirmed blocking all traffic from DMZ to Corporate LAN"

    # Apply Data Validations
    validations["test_result"].add(f"G{start_row+1}:G{end_row}")
    ws.add_data_validation(validations["test_result"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Pass"'], fill=styles["pass_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Fail"'], fill=styles["fail_fill"]["fill"])
    )
    
    # Testing Methods
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "Segmentation Testing Methods (per IMP-S5)"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 12):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    ws["A" + str(row)] = "Test Method"
    ws["B" + str(row)] = "Description"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    methods = [
        ("Ping Test", "Simple connectivity test - should fail if segmentation working"),
        ("Traceroute", "Verify traffic path - should stop at firewall/router"),
        ("Port Scan (nmap)", "Verify ports blocked between zones"),
        ("HTTP/HTTPS Request", "Test application layer connectivity"),
        ("Packet Capture", "Capture traffic to verify deny/drop behaviour"),
        ("Lateral Movement Test", "Attempt to pivot from one zone to another (pentesting)"),
    ]
    
    row += 1
    for method, description in methods:
        ws[f"A{row}"] = method
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"] = description
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [15, 25, 18, 18, 15, 15, 12, 20, 12, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 12: SHEET 8 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(ws, styles, validations):
    """Create Gap Analysis sheet for segmentation issues."""
    
    ws.title = "Gap Analysis"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "SEGMENTATION GAPS - REMEDIATION TRACKING"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document segmentation gaps: flat networks, excessive trust, missing controls, bypass routes."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",               # A - Auto
        "Gap Type",             # B (Flat Network, Excessive Trust, etc.)
        "Location/Zone",        # C
        "Description",          # D
        "Risk",                 # E
        "Severity",             # F - Dropdown
        "Remediation Plan",     # G
        "Owner",                # H
        "Status",               # I - Dropdown
        "Target Date",          # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1
    
    _sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx in range(start_row, end_row + 1):
        # Gap ID
        _c1 = ws.cell(row=row_idx, column=1)
        if row_idx == start_row:
            _c1.value = "SEG-GAP-001"
            _c1.fill = _sample_fill
            _c1.border = _sample_border
            _c1.font = Font(color="808080", name="Calibri")
        else:
            gap_num = row_idx - start_row + 1
            _c1.value = f'=IF(B{row_idx}<>"","SEG-GAP-" & TEXT({gap_num},"000"),"")'
            _c1.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            _c1.border = _sample_border

        # Input cells
        for col in range(2, 11):
            cell = ws.cell(row=row_idx, column=col)
            if row_idx == start_row:
                cell.fill = _sample_fill
                cell.border = _sample_border
                cell.font = Font(color="808080", name="Calibri")
            else:
                apply_style(cell, styles["input_cell"])

    # Sample data for row 4
    ws.cell(row=start_row, column=1).value = "SEG-GAP-001"
    ws.cell(row=start_row, column=2).value = "Flat Network"
    ws.cell(row=start_row, column=3).value = "Server VLAN (10.10.2.0/24)"
    ws.cell(row=start_row, column=4).value = "No segmentation between application and database server tiers"
    ws.cell(row=start_row, column=5).value = "Lateral movement possible from app to database servers"
    ws.cell(row=start_row, column=6).value = "Critical"
    ws.cell(row=start_row, column=7).value = "Create dedicated database VLAN; permit only app\u2192DB on port 1433 via firewall"
    ws.cell(row=start_row, column=8).value = "Network Architecture Team"
    ws.cell(row=start_row, column=9).value = "Open"
    ws.cell(row=start_row, column=10).value = "28.02.2026"

    # Apply Data Validations
    validations["gap_severity"].add(f"F{start_row+1}:F{end_row}")
    ws.add_data_validation(validations["gap_severity"])

    validations["gap_status"].add(f"I{start_row+1}:I{end_row}")
    ws.add_data_validation(validations["gap_status"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Common Gap Types
    row = end_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "Common Segmentation Gaps"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 11):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    gap_types = [
        "Flat Network: No segmentation; all devices in same broadcast domain (Critical)",
        "Excessive Trust: Guest/DMZ can access internal resources (High)",
        "Missing Firewall Rules: Zones defined but not enforced by firewall (High)",
        "Overly Permissive Rules: 'any any any' rules bypass segmentation (Medium)",
        "VLAN Hopping Risk: Native VLAN 1, DTP enabled (Medium)",
        "Bypass Routes: Alternative paths circumvent segmentation (Critical)",
        "Outdated Rules: Firewall rules not reviewed, unused rules accumulate (Low)",
    ]
    
    for gap in gap_types:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = gap
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    widths = [15, 25, 20, 40, 35, 12, 40, 20, 15, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 13: SHEET 9 - SUMMARY DASHBOARD
# ============================================================================

def create_compliance_summary_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1, TABLE 2, TABLE 3."""

    ws.title = "Summary Dashboard"

    SQ = "'"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _white_bold_14 = Font(bold=True, color="FFFFFF", size=14)
    _white_bold_11 = Font(bold=True, color="FFFFFF", size=11)
    _dark_bold_10 = Font(bold=True, color="000000", size=10)
    _dark_10 = Font(bold=False, color="000000", size=10)
    _dark_9_italic = Font(bold=False, color="000000", size=9, italic=True)
    _title_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _data_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _t3total_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _t3_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    _center = Alignment(horizontal="center", vertical="center")
    _left = Alignment(horizontal="left", vertical="center")
    _wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _banner(row, text, font, fill):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = font
        c.fill = fill

    def _hdr_row_t1(row):
        for col_letter, text in [("A", "Assessment Area"), ("B", "Total"), ("C", "Yes"), ("D", "Partial"), ("E", "No"), ("F", "N-A"), ("G", "Compliance %")]:
            c = ws[f"{col_letter}{row}"]
            c.value = text
            c.font = _dark_bold_10
            c.fill = _hdr_fill
            c.border = _thin
            c.alignment = _center

    def _data_row(row, area, b, c_val, d, e, f_val, g):
        ws[f"A{row}"].value = area
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col, val in [("B", b), ("C", c_val), ("D", d), ("E", e), ("F", f_val)]:
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.font = _dark_10
            cell.border = _thin
            cell.alignment = _center
        ws[f"G{row}"].value = g
        ws[f"G{row}"].font = _dark_10
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    def _total_row_t1(row, first, last):
        ws[f"A{row}"].value = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"A{row}"].fill = _total_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        for col in ["B", "C", "D", "E", "F"]:
            c = ws[f"{col}{row}"]
            c.value = f"=SUM({col}{first}:{col}{last})"
            c.font = Font(bold=True, color="000000", size=10)
            c.fill = _total_fill
            c.border = _thin
            c.alignment = _center
        ws[f"G{row}"].value = f"=IF((B{row}-F{row})=0,\"0%\",ROUND(C{row}/(B{row}-F{row})*100,1)&\"%\")"
        ws[f"G{row}"].font = Font(bold=True, color="000000", size=10)
        ws[f"G{row}"].fill = _total_fill
        ws[f"G{row}"].border = _thin
        ws[f"G{row}"].alignment = _center

    # ── ROW 1: Title ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 35
    _banner(1, "NETWORK SEGMENTATION MATRIX & ASSESSMENT — SUMMARY DASHBOARD", _white_bold_14, _title_fill)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── ROW 2: Subtitle ───────────────────────────────────────────────
    ws["A2"] = "ISO/IEC 27001:2022 — Controls A.8.20 · A.8.21 · A.8.22: Network Security, Security of Network Services and Segregation of Networks"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = _left
    ws.merge_cells("A2:G2")

    # ── TABLE 1 ───────────────────────────────────────────────────────
    _banner(4, "TABLE 1: ASSESSMENT AREA COMPLIANCE", _white_bold_11, _title_fill)
    _hdr_row_t1(5)

    # Row 6: Segmentation Testing (col G Test Result, rows 4-33)
    # DV values: Pass / Fail / Partial / Not Tested
    _data_row(6,
        "Segmentation Testing",
        "=COUNTA('Segmentation Testing'!B5:B33)",
        "=COUNTIF('Segmentation Testing'!G5:G33,\"Pass\")",
        "=COUNTIF('Segmentation Testing'!G5:G33,\"Partial\")",
        "=COUNTIF('Segmentation Testing'!G5:G33,\"Fail\")",
        "=B6-(C6+D6+E6)",
        "=IF((B6-F6)=0,\"0%\",ROUND(C6/(B6-F6)*100,1)&\"%\")"
    )

    # Row 7: Firewall Rules Assessment (col J Review Status, rows 4-103)
    # DV values: Current / Outdated / Unused / Requires Review
    _data_row(7,
        "Firewall Rules Assessment",
        "=COUNTA('Firewall Rules Assessment'!B5:B103)",
        "=COUNTIF('Firewall Rules Assessment'!J5:J103,\"Current\")",
        "=COUNTIF('Firewall Rules Assessment'!J5:J103,\"Requires Review\")",
        "=COUNTIF('Firewall Rules Assessment'!J5:J103,\"Outdated\")+COUNTIF('Firewall Rules Assessment'!J5:J103,\"Unused\")",
        "=B7-(C7+D7+E7)",
        "=IF((B7-F7)=0,\"0%\",ROUND(C7/(B7-F7)*100,1)&\"%\")"
    )

    # Row 8: Gap Analysis (col I status, rows 4-53)
    # DV values: Open / In Progress / Resolved / Accepted Risk
    _data_row(8,
        "Gap Analysis",
        "=COUNTA('Gap Analysis'!B5:B53)",
        "=COUNTIF('Gap Analysis'!I5:I53,\"Resolved\")",
        "=COUNTIF('Gap Analysis'!I5:I53,\"In Progress\")",
        "=COUNTIF('Gap Analysis'!I5:I53,\"Open\")",
        "=B8-(C8+D8+E8)",
        "=IF((B8-F8)=0,\"0%\",ROUND(C8/(B8-F8)*100,1)&\"%\")"
    )

    # Row 9: TOTAL
    _total_row_t1(9, 6, 8)

    # ── TABLE 2 ───────────────────────────────────────────────────────
    _banner(11, "TABLE 2: KEY METRICS", _white_bold_11, _title_fill)

    ws["A12"].value = "Metric"
    ws["A12"].font = _dark_bold_10
    ws["A12"].fill = _hdr_fill
    ws["A12"].border = _thin
    ws["A12"].alignment = _left
    ws["B12"].value = "Value"
    ws["B12"].font = _dark_bold_10
    ws["B12"].fill = _hdr_fill
    ws["B12"].border = _thin
    ws["B12"].alignment = _center
    ws.merge_cells("C12:G12")
    ws["C12"].value = "What This Shows"
    ws["C12"].font = _dark_bold_10
    ws["C12"].fill = _hdr_fill
    ws["C12"].border = _thin
    ws["C12"].alignment = _left

    t2_metrics = [
        (13, "Failed Segmentation Tests",
         "=COUNTIF('Segmentation Testing'!G5:G33,\"Fail\")",
         "Tests confirming bypass routes or misconfigurations (immediate investigation required)"),
        (14, "Untested Segments",
         "=COUNTIF('Segmentation Testing'!G5:G33,\"Not Tested\")",
         "Test cases not yet executed (unknown segmentation effectiveness)"),
        (15, "Partial Test Results",
         "=COUNTIF('Segmentation Testing'!G5:G33,\"Partial\")",
         "Tests with partial pass (incomplete segmentation enforcement)"),
        (16, "Outdated Firewall Rules",
         "=COUNTIF('Firewall Rules Assessment'!J5:J103,\"Outdated\")",
         "Firewall rules requiring immediate review (may permit unauthorised traffic)"),
        (17, "Unused Firewall Rules",
         "=COUNTIF('Firewall Rules Assessment'!J5:J103,\"Unused\")",
         "Unused rules creating unnecessary attack surface (remove or review)"),
        (18, "Rules Requiring Review",
         "=COUNTIF('Firewall Rules Assessment'!J5:J103,\"Requires Review\")",
         "Rules overdue for periodic review (potential policy violations)"),
        (19, "Open Segmentation Gaps",
         "=COUNTIF('Gap Analysis'!I5:I53,\"Open\")",
         "Outstanding segmentation gaps without active remediation"),
        (20, "Overall Compliance Rate",
         "=G9",
         "Overall network segmentation compliance percentage (TABLE 1 total)"),
    ]

    for row, metric, formula, description in t2_metrics:
        ws[f"A{row}"].value = metric
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = description
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).border = _thin

    # ── TABLE 3 ───────────────────────────────────────────────────────
    _banner(22, "TABLE 3: CRITICAL FINDINGS", _white_bold_11, _t3_fill)

    ws["A23"].value = "Critical Finding Type"
    ws["A23"].font = _dark_bold_10
    ws["A23"].fill = _hdr_fill
    ws["A23"].border = _thin
    ws["A23"].alignment = _left
    ws["B23"].value = "Count"
    ws["B23"].font = _dark_bold_10
    ws["B23"].fill = _hdr_fill
    ws["B23"].border = _thin
    ws["B23"].alignment = _center
    ws.merge_cells("C23:G23")
    ws["C23"].value = "Filter Instructions"
    ws["C23"].font = _dark_bold_10
    ws["C23"].fill = _hdr_fill
    ws["C23"].border = _thin
    ws["C23"].alignment = _left

    t3_findings = [
        (24, "Failed Segmentation Tests",
         "=COUNTIF('Segmentation Testing'!G5:G33,\"Fail\")",
         "Filter Segmentation Testing: Test Result = \"Fail\""),
        (25, "Outdated Firewall Rules",
         "=COUNTIF('Firewall Rules Assessment'!J5:J103,\"Outdated\")",
         "Filter Firewall Rules Assessment: Review Status = \"Outdated\""),
        (26, "Open Segmentation Gaps",
         "=COUNTIF('Gap Analysis'!I5:I53,\"Open\")",
         "Filter Gap Analysis: Status = \"Open\""),
        (27, "Untested Segments",
         "=COUNTIF('Segmentation Testing'!G5:G33,\"Not Tested\")",
         "Filter Segmentation Testing: Test Result = \"Not Tested\""),
        (28, "Critical Severity Gaps",
         "=COUNTIF('Gap Analysis'!F5:F53,\"Critical\")",
         "Filter Gap Analysis: Severity = \"Critical\""),
    ]

    for row, finding_type, formula, instructions in t3_findings:
        ws[f"A{row}"].value = finding_type
        ws[f"A{row}"].font = _dark_10
        ws[f"A{row}"].fill = _data_fill
        ws[f"A{row}"].border = _thin
        ws[f"A{row}"].alignment = _left
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = _dark_10
        ws[f"B{row}"].fill = _data_fill
        ws[f"B{row}"].border = _thin
        ws[f"B{row}"].alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        ws[f"C{row}"].value = instructions
        ws[f"C{row}"].font = _dark_9_italic
        ws[f"C{row}"].fill = _data_fill
        ws[f"C{row}"].border = _thin
        ws[f"C{row}"].alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).fill = _data_fill
            ws.cell(row=row, column=_c).border = _thin

    # TOTAL row
    ws["A29"].value = "TOTAL"
    ws["A29"].font = Font(bold=True, color="000000", size=10)
    ws["A29"].border = _thin
    ws["A29"].alignment = _left
    ws["B29"].value = "=SUM(B24:B28)"
    ws["B29"].font = Font(bold=True, color="000000", size=10)
    ws["B29"].fill = _t3total_fill
    ws["B29"].border = _thin
    ws["B29"].alignment = _center
    ws.merge_cells("C29:G29")
    ws["C29"].value = "Total critical findings requiring immediate remediation"
    ws["C29"].font = Font(italic=True, size=9, color="000000")
    ws["C29"].alignment = _left

    # ── Column widths ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 14: SHEET 10 - REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap_sheet(ws, styles, validations):
    """Create Remediation Roadmap for segmentation improvements."""
    
    ws.title = "Remediation Roadmap"
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "SEGMENTATION REMEDIATION ROADMAP"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "Prioritised plan for segmentation improvements. Focus on Critical gaps first."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Priority",             # A
        "Severity",             # B - Dropdown
        "Improvement",          # C
        "Affected Zones",       # D
        "Remediation Action",   # E
        "Owner",                # F
        "Target Date",          # G
        "Status",               # H - Dropdown
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + REMEDIATION_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Priority (manual)
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        
        # Input cells
        for col in range(2, 9):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["gap_severity"].add(f"B{start_row}:B{end_row}")
    ws.add_data_validation(validations["gap_severity"])
    
    validations["gap_status"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["gap_status"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    
    # Remediation Approach
    row = end_row + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Segmentation Remediation Approach"
    apply_style(ws[f"A{row}"], styles["header"])
    for _c in range(1, 9):
        ws.cell(row=row, column=_c).border = styles["header"]["border"]

    row += 1
    approach = [
        "1. CRITICAL FIRST: Address Critical gaps immediately (flat networks, bypass routes)",
        "2. QUICK WINS: Implement easy improvements (VLAN assignment, ACL updates)",
        "3. FIREWALL RULES: Review and tighten firewall rules between zones",
        "4. TESTING: Re-test after each change to verify effectiveness",
        "5. DOCUMENTATION: Update network documentation (topology, VLAN database)",
        "6. MONITORING: Enable logging and alerting for inter-zone traffic",
    ]
    
    for item in approach:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = item
        apply_style(ws[f"A{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    
    ws.freeze_panes = "A4"




# ============================================================================
# EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (italic, NOT blue banner) ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers (003366, white text) ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Network Segmentation Assessment",
        3: "Screenshot",
        4: "Network segmentation test results and firewall rule review documentation",
        5: "/evidence/A.8.20-22/",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"

# ============================================================================
# APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws.title = "Approval Sign-Off"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE(\'Summary Dashboard\'!G6:G8),\"\")")  ,
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    styles = _STYLES
    validations = create_data_validations()

    create_instructions_sheet(wb.create_sheet())
    create_security_zones_sheet(wb.create_sheet(), styles, validations)
    create_segmentation_matrix_sheet(wb.create_sheet(), styles, validations)
    create_vlan_inventory_sheet(wb.create_sheet(), styles)
    create_firewall_rules_sheet(wb.create_sheet(), styles, validations)
    create_acl_assessment_sheet(wb.create_sheet(), styles, validations)
    create_segmentation_testing_sheet(wb.create_sheet(), styles, validations)
    create_gap_analysis_sheet(wb.create_sheet(), styles, validations)
    create_remediation_roadmap_sheet(wb.create_sheet(), styles, validations)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    create_compliance_summary_sheet(wb.create_sheet(), styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_sheet(ws)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path.name}")

def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"ERROR saving workbook: {e}")
        sys.exit(1)
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\nAssessment Sheets:")
    logger.info(f"  • Security Zones Inventory ({ZONE_ROW_COUNT} zones)")
    logger.info("  • Segmentation Matrix (zone-to-zone traffic rules)")
    logger.info(f"  • VLAN Inventory ({VLAN_ROW_COUNT} VLANs)")
    logger.info(f"  • Firewall Rules Assessment ({FIREWALL_RULE_COUNT} rules)")
    logger.info(f"  • ACL Assessment ({ACL_COUNT} ACLs)")
    logger.info(f"  • Segmentation Testing ({TEST_ROW_COUNT} effectiveness tests)")
    logger.info("\n Analysis & Reporting:")
    logger.info(f"  • Gap Analysis ({GAP_ROW_COUNT} gap tracking)")
    logger.info("  • Summary Dashboard (metrics, pie chart)")
    logger.info(f"  • Remediation Roadmap ({REMEDIATION_ROW_COUNT} prioritised actions)")
    logger.info("\n Common Sheets:")
    logger.info("  • Evidence Register (100 evidence rows)")
    logger.info("  • Approval Sign-Off (multi-stakeholder workflow)")
    logger.info("\n" + "=" * 80)
    logger.info("NEXT STEPS:")
    logger.info("  1. Define all security zones in Security Zones Inventory")
    logger.info("  2. Complete Segmentation Matrix (zone-to-zone rules)")
    logger.info("  3. Document VLANs and firewall rules")
    logger.info("  4. Test segmentation effectiveness")
    logger.info("  5. Identify and remediate gaps")
    logger.info("\n" + "=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
