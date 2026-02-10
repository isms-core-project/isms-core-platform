#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.9.5 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Consolidated Compliance Dashboard - All Four Assessment Domains

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific reporting requirements, stakeholder needs, and
dashboard visualization preferences.

Key customization areas:
1. Executive dashboard metrics (match your KPI framework)
2. External workbook references (adapt to your file naming conventions)
3. Consolidation formulas (align with your data structures)
4. Visualization preferences (charts, conditional formatting)
5. Stakeholder approval workflow (based on your governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel compliance dashboard that consolidates
data from all four configuration management assessment domains (A.8.9.1 through
A.8.9.4) to provide executive oversight and audit readiness for ISO 27001:2022
Control A.8.9.

**Purpose:**
Enables consolidated view of configuration management maturity across baseline,
change control, monitoring, and hardening domains with aggregated compliance
metrics, gap analysis, risk assessment, and remediation tracking.

**Assessment Scope:**
- Executive dashboard with consolidated metrics
- Gap analysis across all four domains
- Risk register with aggregated risk assessment
- Remediation roadmap with prioritization
- KPIs and metrics tracking
- Evidence register consolidation
- Action items and follow-up tracking
- Audit and compliance log
- Multi-stakeholder approval workflow

**Generated Workbook Structure:**
1. Executive Dashboard - Consolidated metrics with external workbook links
2. Gap Analysis - Consolidated gaps from all 4 domains (200 entries)
3. Risk Register - Aggregated risk assessments (100 entries)
4. Remediation Roadmap - Prioritized remediation actions (200 entries)
5. KPIs & Metrics - Performance tracking across all domains (50+ KPIs)
6. Evidence Register - Consolidated audit evidence (500 entries)
7. Action Items & Follow-up - Tracked action items (200 entries)
8. Audit & Compliance Log - Detailed compliance tracking (100 entries)
9. Approval & Sign-Off - Executive approval workflow

**Key Features:**
- External workbook formula links to source assessments
- Automated metric consolidation and aggregation
- Conditional formatting for compliance status
- Risk-based prioritization for remediation
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Executive-level approval workflow

**Integration:**
This dashboard consolidates data from:
    - ISMS-IMP-A.8.9.1.xlsx (Baseline Configuration)
    - ISMS-IMP-A.8.9.2.xlsx (Change Control)
    - ISMS-IMP-A.8.9.3.xlsx (Configuration Monitoring)
    - ISMS-IMP-A.8.9.4.xlsx (Security Hardening)

**CRITICAL SETUP REQUIREMENT:**
This dashboard uses external workbook formula references. All source assessment
workbooks MUST be in the same directory as the dashboard. After generation:
1. Run normalize_assessment_files_a89.py to standardize source filenames
2. Place dashboard in same folder as normalized source files
3. Open dashboard and click "Update Links" when prompted
4. #REF errors before updating links are EXPECTED and will resolve

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_5_compliance_dashboard.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_5_compliance_dashboard.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_5_compliance_dashboard.py --date 20250127

Output:
    File: ISMS_IMP_A_8_9_5_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Run normalize_assessment_files_a89.py on source assessments
    2. Place dashboard in same directory as normalized source files
    3. Open dashboard in Excel
    4. Click "Update Links" when prompted (REQUIRED)
    5. Verify all external references resolve (no #REF errors)
    6. Review consolidated metrics on Executive Dashboard
    7. Validate gap analysis consolidation
    8. Review risk register and prioritization
    9. Update remediation roadmap with timelines
    10. Collect consolidated evidence references
    11. Obtain executive approvals
    12. Distribute to stakeholders

Alternative Data Consolidation:
    If external workbook links don't work in your environment, use:
    python3 consolidate_a89_dashboard.py ISMS_IMP_A_8_9_5_Dashboard.xlsx
    
    This alternative method copies data values instead of using formula links.

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Dashboard Type:       Consolidated Compliance Dashboard (All 4 Domains)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.1: Baseline Configuration Assessment (Domain 1)
    - ISMS-IMP-A.8.9.2: Change Control Assessment (Domain 2)
    - ISMS-IMP-A.8.9.3: Configuration Monitoring Assessment (Domain 3)
    - ISMS-IMP-A.8.9.4: Security Hardening Assessment (Domain 4)

Related Scripts:
    - generate_a89_1_baseline_configuration.py
    - generate_a89_2_change_control.py
    - generate_a89_3_monitoring.py
    - generate_a89_4_hardening.py
    - normalize_assessment_files_a89.py
    - consolidate_a89_dashboard.py (alternative consolidation method)
    - excel_sanity_check_a89.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements consolidated dashboard per ISMS-IMP-A.8.9.5 specification
    - External workbook formula linking to all 4 source domains
    - Integrated metric aggregation and gap consolidation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**External Workbook Links - CRITICAL:**
This dashboard REQUIRES external workbook formula links to function properly.
The Executive Dashboard contains formulas like:
    ='[ISMS-IMP-A.8.9.1.xlsx]Gap_Analysis'!$B$15

These formulas will show #REF errors until:
1. Source workbooks are in the same directory as dashboard
2. Excel's "Update Links" is clicked when opening dashboard
3. Source workbook filenames match exactly (use normalize script)

If external links don't work in your environment (SharePoint, email distribution,
security restrictions), use the alternative consolidation script:
    python3 consolidate_a89_dashboard.py ISMS_IMP_A_8_9_5_Dashboard.xlsx

**Audit Considerations:**
This dashboard consolidates audit evidence from all four assessment domains.
Ensure source assessments are complete and approved before dashboard review.
Auditors will use this dashboard as entry point for detailed evidence review.

**Data Protection:**
Dashboard contains highly sensitive aggregated information including:
- Consolidated security gaps across entire infrastructure
- Risk assessments and vulnerability details
- Remediation priorities and timelines
- Executive compliance status

Handle in accordance with your organization's highest data classification.

**Maintenance:**
Update dashboard:
- Monthly: Refresh external links to get latest assessment data
- Quarterly: Complete review with executive stakeholders
- Annually: Archive dashboard snapshot for historical trending
- Ad-hoc: When source assessments are significantly updated

**Quality Assurance:**
Have Configuration Manager and CISO validate dashboard before executive
presentation. Verify all external links resolve and metrics are accurate.

**Regulatory Alignment:**
Dashboard supports compliance reporting for:
- ISO 27001:2022: Control A.8.9 evidence consolidation
- Executive reporting: Board and C-level compliance oversight
- Audit readiness: Consolidated evidence for external audits
- Risk management: Aggregated risk view for decision-making

Customize dashboard metrics to align with organizational KPI framework.

**Troubleshooting External Links:**
If #REF errors persist after "Update Links":
1. Verify all source files in same directory as dashboard
2. Check filenames match exactly (case-sensitive)
3. Ensure source files are not open in another Excel instance
4. Try "Edit Links" > "Update Values" > "Close"
5. If still failing, use consolidate_a89_dashboard.py script

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
from datetime import datetime, timedelta
from typing import List, Dict, Tuple

# =============================================================================
# Third-Party Imports
# =============================================================================
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.9.5"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.8.9"
CONTROL_NAME = "Configuration Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

# =============================================================================
# CONFIGURATION SECTION - CUSTOMIZE FOR YOUR ORGANIZATION
# =============================================================================

# CUSTOMIZE: Dashboard metadata
CONFIG = {
    'organization_name': '[Organisation]',
    'dashboard_date': datetime.now().strftime('%d.%m.%Y'),
    'dashboard_owner': '[ISMS Manager]',
    'output_filename': f'ISMS_A_8_9_5_Compliance_Dashboard_{datetime.now().strftime("%Y%m%d")}.xlsx',
}

# CUSTOMIZE: Domain compliance weights (must sum to 1.0)
DOMAIN_WEIGHTS = {
    'baseline_configuration': 0.20,    # 20% - Foundation for management
    'change_control': 0.25,            # 25% - Highest operational risk
    'configuration_monitoring': 0.20,  # 20% - Detection capability
    'security_hardening': 0.25,        # 25% - Security criticality
    'process_maturity': 0.10,          # 10% - Governance sustainability
}

# Verify weights sum to 100%
assert abs(sum(DOMAIN_WEIGHTS.values()) - 1.0) < 0.01, "Domain weights must sum to 1.0"

# CUSTOMIZE: Compliance targets
COMPLIANCE_TARGETS = {
    'overall': 95,              # Overall A.8.9 target
    'domain_minimum': 90,       # Minimum per domain
    'critical_gaps': 0,         # Target for critical gaps
    'evidence_completeness': 95, # Evidence documentation target
}

# Normalized source workbook filenames (NO DATES)
SOURCE_WORKBOOKS = {
    'baseline': 'ISMS-IMP-A.8.9.1.xlsx',
    'change': 'ISMS-IMP-A.8.9.2.xlsx',
    'monitoring': 'ISMS-IMP-A.8.9.3.xlsx',
    'hardening': 'ISMS-IMP-A.8.9.4.xlsx',
}

# CUSTOMIZE: External workbook cell references
# These specify which cells to pull from each source workbook
EXTERNAL_REFERENCES = {
    'baseline': {
        'compliance_percentage': 'Compliance_Dashboard!B5',
        'total_assets': 'Asset_Inventory!B2',
        'documented_baselines': 'Compliance_Dashboard!B8',
        'version_control_compliance': 'Compliance_Dashboard!B12',
        'critical_gaps': 'Compliance_Dashboard!B15',
        'last_assessment_date': 'Approval_Sign_Off!B6',
    },
    'change': {
        'compliance_percentage': 'Change_Summary!B10',
        'total_changes': 'Change_Summary!B5',
        'emergency_changes': 'Change_Summary!B14',
        'change_success_rate': 'Change_Summary!B18',
        'critical_gaps': 'Change_Summary!B22',
        'last_assessment_date': 'Approval_Sign_Off!B6',
    },
    'monitoring': {
        'compliance_percentage': 'Compliance_Dashboard!B8',
        'monitoring_coverage': 'Compliance_Dashboard!B12',
        'mttd_hours': 'Compliance_Dashboard!B16',
        'critical_drift_count': 'Compliance_Dashboard!B20',
        'critical_gaps': 'Compliance_Dashboard!B24',
        'last_assessment_date': 'Approval_Sign_Off!B6',
    },
    'hardening': {
        'compliance_percentage': 'Compliance_Dashboard!B5',
        'total_controls': 'Compliance_Dashboard!B8',
        'high_risk_gaps': 'Compliance_Dashboard!B13',
        'exception_rate': 'Compliance_Dashboard!B15',
        'critical_gaps': 'Compliance_Dashboard!B14',
        'last_assessment_date': 'Approval_Sign_Off!B6',
    },
}

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

# Fonts
FONT_HEADER = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='Calibri', size=18, bold=True)
FONT_SUBHEADER = Font(name='Calibri', size=12, bold=True)
FONT_NORMAL = Font(name='Calibri', size=11)
FONT_SMALL = Font(name='Calibri', size=10)
FONT_BOLD = Font(name='Calibri', size=11, bold=True)
FONT_LARGE = Font(name='Calibri', size=24, bold=True)

# Fills
FILL_HEADER = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
FILL_SUBHEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
FILL_LIGHT_BLUE = PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid')
FILL_LIGHT_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_LIGHT_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_LIGHT_RED = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
FILL_GREEN = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
FILL_DARK_GREEN = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_RED = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
FILL_ORANGE = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
FILL_GRAY = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

THICK_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def set_column_widths(ws, widths: Dict[str, float]):
    """Set column widths for a worksheet."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_header_row(ws, row: int, headers: List[str], start_col: int = 1):
    """Apply header formatting to a row."""
    for idx, header in enumerate(headers, start=start_col):
        col_letter = get_column_letter(idx)
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def create_external_reference(workbook_name: str, sheet_name: str, cell_ref: str) -> str:
    """
    Create Excel external workbook reference (for use inside formulas).

    Args:
        workbook_name: Name of source workbook (e.g., 'ISMS-IMP-A.8.9.1.xlsx')
        sheet_name: Sheet name in source workbook
        cell_ref: Cell reference (e.g., 'B5')

    Returns:
        External reference string (without leading =, for use in formulas)
    """
    # External references always need quotes around workbook/sheet path
    return f"'[{workbook_name}]{sheet_name}'!{cell_ref}"


def protect_sheet(ws, password: str = None):
    """Protect worksheet while allowing data entry in unprotected cells."""
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


def unlock_cell_range(ws, cell_range: str):
    """Unlock a range of cells for data entry."""
    for row in ws[cell_range]:
        for cell in row:
            cell.protection = Protection(locked=False)

# =============================================================================
# SHEET GENERATION FUNCTIONS - PART 1
# =============================================================================

def create_instructions_sheet(wb: Workbook) -> None:
    """Create the Instructions sheet."""
    ws = wb.create_sheet("Instructions", 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100
    
    row = 1
    
    # Title
    ws[f'A{row}'] = 'ISMS Control A.8.9.5'
    ws[f'A{row}'].font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws[f'A{row}'] = 'Compliance Dashboard - Configuration Management'
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells(f'A{row}:B{row}')
    row += 2
    
    # Instructions sections
    instructions = [
        ("PURPOSE", [
            "This dashboard consolidates compliance data from all four A.8.9 domain assessments "
            "into a single executive-level view of overall configuration management posture.",
            "",
            "Dashboard aggregates:",
            "\u2022 A.8.9.1 - Baseline Configuration Assessment",
            "\u2022 A.8.9.2 - Change Control Assessment",
            "\u2022 A.8.9.3 - Configuration Monitoring Assessment",
            "\u2022 A.8.9.4 - Security Hardening Assessment",
        ]),
        
        ("KEY CONCEPTS", [
            "External Workbook Linking: Dashboard formulas reference other Excel files in the "
            "same directory using formulas like ='[ISMS-IMP-A.8.9.1.xlsx]Sheet'!Cell",
            "",
            "Normalized Filenames: Source files must have date-free names (ISMS-IMP-A.8.9.X.xlsx) "
            "created by running normalize_assessment_files_a89.py",
            "",
            "Overall Compliance: Weighted average across 5 domains:",
            f"\u2022 Baseline Configuration: {int(DOMAIN_WEIGHTS['baseline_configuration']*100)}%",
            f"\u2022 Change Control: {int(DOMAIN_WEIGHTS['change_control']*100)}%",
            f"\u2022 Configuration Monitoring: {int(DOMAIN_WEIGHTS['configuration_monitoring']*100)}%",
            f"\u2022 Security Hardening: {int(DOMAIN_WEIGHTS['security_hardening']*100)}%",
            f"\u2022 Process Maturity: {int(DOMAIN_WEIGHTS['process_maturity']*100)}%",
            "",
            f"Target: ≥{COMPLIANCE_TARGETS['overall']}% overall compliance",
        ]),
        
        ("SETUP WORKFLOW", [
            "1. Complete all 4 domain assessments (A.8.9.1 through A.8.9.4)",
            "2. Generate assessment workbooks with Python scripts",
            "3. Run normalization: python3 normalize_assessment_files_a89.py",
            "4. Verify normalized files exist in Dashboard_Sources directory:",
            "   \u2022 ISMS-IMP-A.8.9.1.xlsx",
            "   \u2022 ISMS-IMP-A.8.9.2.xlsx",
            "   \u2022 ISMS-IMP-A.8.9.3.xlsx",
            "   \u2022 ISMS-IMP-A.8.9.4.xlsx",
            "5. Place this dashboard in same directory as normalized files",
            "6. Open dashboard and click 'Update Links' when Excel prompts",
            "7. Verify Workbook_Integration_Settings shows all 4 workbooks linked",
        ]),
        
        ("USING THE DASHBOARD", [
            "Overall_Compliance_Dashboard: Primary executive view, single-page summary",
            "",
            "Asset_Compliance_Summary: Per-asset view across all domains",
            "",
            "Gap_Prioritization: Consolidated gap list with risk ranking",
            "",
            "Trend_Analysis: Historical compliance tracking (requires multiple cycles)",
            "",
            "Evidence_Register: Links to domain-specific evidence",
            "",
            "Approval_Sign_Off: Executive approval and governance",
        ]),
        
        ("EXTERNAL WORKBOOK LINKING", [
            "How it works:",
            "\u2022 Dashboard formulas reference source workbooks automatically",
            "\u2022 Excel updates when source workbooks change",
            "\u2022 Links are relative - files must be in same directory",
            "\u2022 If workbook not found: Formula shows #REF! error",
            "",
            "Troubleshooting broken links:",
            "\u2022 Verify all 4 normalized workbooks exist in same directory",
            "\u2022 Excel → Data → Edit Links → Update Values",
            "\u2022 Check filenames match exactly (case-sensitive)",
            "\u2022 Ensure workbooks are not open in exclusive mode",
        ]),
        
        ("UPDATE FREQUENCY", [
            "Domain Assessments: Quarterly minimum",
            "Dashboard: Updates automatically when sources change",
            "Executive Review: Monthly minimum",
            "Trend Analysis: Requires 3+ assessment cycles",
        ]),
        
        ("ROLES & RESPONSIBILITIES", [
            "Dashboard Owner: Coordinates assessment schedule, ensures currency",
            "Domain Assessors: Complete domain-specific assessments per schedule",
            "Executive Sponsor: Reviews dashboard, approves remediation priorities",
            "ISMS Manager: Oversees overall A.8.9 compliance",
            "Auditor: Verifies dashboard accuracy, validates evidence links",
        ]),
    ]
    
    for section_title, section_content in instructions:
        # Section header
        ws[f'A{row}'] = section_title
        ws[f'A{row}'].font = FONT_SUBHEADER
        ws[f'A{row}'].fill = FILL_SUBHEADER
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # Section content
        for line in section_content:
            ws[f'B{row}'] = line
            ws[f'B{row}'].font = FONT_NORMAL
            ws[f'B{row}'].alignment = ALIGN_LEFT
            row += 1
        
        row += 1  # Blank line between sections
    
    # Document info at bottom
    row += 2
    ws[f'A{row}'] = 'Document ID:'
    ws[f'B{row}'] = 'ISMS-IMP-A.8.9.5-Dashboard'
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    ws[f'A{row}'] = 'Version:'
    ws[f'B{row}'] = '1.0'
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    ws[f'A{row}'] = 'Date:'
    ws[f'B{row}'] = CONFIG['dashboard_date']
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    ws[f'A{row}'] = 'Dashboard Owner:'
    ws[f'B{row}'] = CONFIG['dashboard_owner']
    ws[f'A{row}'].font = FONT_BOLD


def create_workbook_integration_settings(wb: Workbook) -> None:
    """Create the Workbook_Integration_Settings sheet."""
    ws = wb.create_sheet("Workbook_Integration_Settings")
    
    # Set column widths
    widths = {
        'A': 25,  # Domain
        'B': 20,  # Document ID
        'C': 30,  # Expected Filename
        'D': 15,  # Status
        'E': 20,  # Last Modified
        'F': 40,  # Link Status
    }
    set_column_widths(ws, widths)
    
    row = 1
    
    # Title
    ws[f'A{row}'] = 'Workbook Integration Settings'
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells(f'A{row}:F{row}')
    row += 2
    
    # Section 1: Source Workbook Configuration
    ws[f'A{row}'] = 'SOURCE WORKBOOK CONFIGURATION'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers
    headers = ['Domain', 'Document ID', 'Expected Filename', 'Status', 'Last Modified', 'Link Status']
    apply_header_row(ws, row, headers)
    row += 1
    
    # Source workbook rows
    sources = [
        ('Baseline Configuration', 'ISMS-IMP-A.8.9.1', SOURCE_WORKBOOKS['baseline']),
        ('Change Control', 'ISMS-IMP-A.8.9.2', SOURCE_WORKBOOKS['change']),
        ('Configuration Monitoring', 'ISMS-IMP-A.8.9.3', SOURCE_WORKBOOKS['monitoring']),
        ('Security Hardening', 'ISMS-IMP-A.8.9.4', SOURCE_WORKBOOKS['hardening']),
    ]
    
    start_row = row
    for domain, doc_id, filename in sources:
        ws[f'A{row}'] = domain
        ws[f'B{row}'] = doc_id
        ws[f'C{row}'] = filename
        
        # Status formula - checks if external reference resolves
        ext_ref = create_external_reference(filename, 'Instructions', 'A1')
        ws[f'D{row}'] = f'=IF(ISERROR({ext_ref}),"Not Found",IF(ISREF({ext_ref}),"Linked","Error"))'
        ws[f'D{row}'].protection = Protection(locked=True)
        
        # Last Modified - try to pull from source (may not work)
        ws[f'E{row}'] = 'N/A'
        
        # Link Status formula
        ws[f'F{row}'] = f'=IF(D{row}="Linked","\u2705 Linked and current",IF(D{row}="Not Found","\u26A0\uFE0F Workbook not found","\u274C Link broken"))'
        ws[f'F{row}'].protection = Protection(locked=True)
        
        row += 1
    
    row += 1
    
    # Section 2: Link Validation Summary
    ws[f'A{row}'] = 'LINK VALIDATION SUMMARY'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    summary_items = [
        ('Total Source Workbooks Expected', '4', '-'),
        ('Workbooks Successfully Linked', f'=COUNTIF(D{start_row}:D{start_row+3},"Linked")', f'=IF(B{row}<4,"Red","Green")'),
        ('Workbooks Missing', f'=COUNTIF(D{start_row}:D{start_row+3},"Not Found")', f'=IF(B{row}>0,"Red","Green")'),
        ('Dashboard Link Health', f'=B{row-1}/B{row-2}', f'=IF(B{row}=1,"Green",IF(B{row}>=0.75,"Yellow","Red"))'),
    ]
    
    for metric, value, status in summary_items:
        ws[f'A{row}'] = metric
        ws[f'A{row}'].font = FONT_BOLD
        if value.startswith('='):
            ws[f'B{row}'] = value
            ws[f'B{row}'].protection = Protection(locked=True)
            if 'Health' in metric:
                ws[f'B{row}'].number_format = '0%'
        else:
            ws[f'B{row}'] = value
        
        if status != '-':
            ws[f'C{row}'] = status
            ws[f'C{row}'].protection = Protection(locked=True)
        
        row += 1
    
    row += 1
    
    # Section 3: Data Pull Test
    ws[f'A{row}'] = 'DATA PULL TEST'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    test_headers = ['Source', 'Test Cell', 'Expected Value Type', 'Actual Value', 'Test Result']
    for idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BORDER
    row += 1
    
    # Test formulas for each source
    test_row = row
    for domain_key, filename in SOURCE_WORKBOOKS.items():
        refs = EXTERNAL_REFERENCES[domain_key]
        sheet_name, cell_ref = refs['compliance_percentage'].split('!')
        
        ws[f'A{row}'] = filename.replace('.xlsx', '')
        ws[f'B{row}'] = refs['compliance_percentage']
        ws[f'C{row}'] = 'Percentage'
        
        ext_ref = create_external_reference(filename, sheet_name, cell_ref)
        ws[f'D{row}'] = f'={ext_ref}'
        ws[f'D{row}'].number_format = '0.0%'
        ws[f'D{row}'].protection = Protection(locked=True)
        
        ws[f'E{row}'] = f'=IF(ISERROR(D{row}),"Fail - Link Broken",IF(AND(D{row}>=0,D{row}<=1),"Pass","Fail - Invalid Data"))'
        ws[f'E{row}'].protection = Protection(locked=True)
        
        row += 1
    
    # Conditional formatting
    # Status column
    ws.conditional_formatting.add(
        f'D{start_row}:D{start_row+3}',
        CellIsRule(operator='equal', formula=['"Linked"'], fill=FILL_GREEN)
    )
    ws.conditional_formatting.add(
        f'D{start_row}:D{start_row+3}',
        CellIsRule(operator='equal', formula=['"Not Found"'], fill=FILL_RED)
    )
    
    # Test Result column
    ws.conditional_formatting.add(
        f'E{test_row}:E{test_row+3}',
        CellIsRule(operator='equal', formula=['"Pass"'], fill=FILL_GREEN)
    )
    ws.conditional_formatting.add(
        f'E{test_row}:E{test_row+3}',
        FormulaRule(formula=[f'LEFT(E{test_row},4)="Fail"'], fill=FILL_RED)
    )


def create_overall_compliance_dashboard(wb: Workbook) -> None:
    """Create the Overall_Compliance_Dashboard sheet."""
    ws = wb.create_sheet("Overall_Compliance_Dashboard")
    
    # Set column widths for dashboard layout
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    row = 1
    
    # Title
    ws[f'A{row}'] = 'Control A.8.9 - Configuration Management'
    ws[f'A{row}'].font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].alignment = ALIGN_CENTER
    row += 1
    
    ws[f'A{row}'] = 'COMPLIANCE DASHBOARD'
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].alignment = ALIGN_CENTER
    row += 1
    
    ws[f'A{row}'] = f'Assessment Date: {CONFIG["dashboard_date"]}'
    ws[f'A{row}'].font = FONT_NORMAL
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].alignment = ALIGN_CENTER
    row += 2
    
    # Section 1: Overall Compliance (Large, prominent)
    ws[f'A{row}'] = 'OVERALL COMPLIANCE'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Overall compliance formula (weighted average)
    compliance_cell_row = row
    ws[f'A{row}'] = 'Overall Compliance Percentage:'
    ws[f'A{row}'].font = FONT_BOLD
    
    # Build weighted average formula
    baseline_ref = create_external_reference(
        SOURCE_WORKBOOKS['baseline'],
        *EXTERNAL_REFERENCES['baseline']['compliance_percentage'].split('!')
    )
    change_ref = create_external_reference(
        SOURCE_WORKBOOKS['change'],
        *EXTERNAL_REFERENCES['change']['compliance_percentage'].split('!')
    )
    monitoring_ref = create_external_reference(
        SOURCE_WORKBOOKS['monitoring'],
        *EXTERNAL_REFERENCES['monitoring']['compliance_percentage'].split('!')
    )
    hardening_ref = create_external_reference(
        SOURCE_WORKBOOKS['hardening'],
        *EXTERNAL_REFERENCES['hardening']['compliance_percentage'].split('!')
    )
    
    # Process maturity calculated from indicators (simplified: average of all domains)
    process_maturity_formula = f'=AVERAGE({baseline_ref},{change_ref},{monitoring_ref},{hardening_ref})'
    
    ws[f'B{row}'] = (
        f'=({baseline_ref}*{DOMAIN_WEIGHTS["baseline_configuration"]}+'
        f'{change_ref}*{DOMAIN_WEIGHTS["change_control"]}+'
        f'{monitoring_ref}*{DOMAIN_WEIGHTS["configuration_monitoring"]}+'
        f'{hardening_ref}*{DOMAIN_WEIGHTS["security_hardening"]}+'
        f'{process_maturity_formula}*{DOMAIN_WEIGHTS["process_maturity"]})'
    )
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = FONT_LARGE
    ws[f'B{row}'].alignment = ALIGN_CENTER
    ws.merge_cells(f'B{row}:C{row}')
    
    ws[f'D{row}'] = 'Target:'
    ws[f'E{row}'] = f'≥{COMPLIANCE_TARGETS["overall"]}%'
    ws[f'E{row}'].alignment = ALIGN_CENTER
    row += 1
    
    # Status indicator
    ws[f'A{row}'] = 'Compliance Status:'
    ws[f'A{row}'].font = FONT_BOLD
    
    # Get critical gaps sum
    baseline_gaps_ref = create_external_reference(
        SOURCE_WORKBOOKS['baseline'],
        *EXTERNAL_REFERENCES['baseline']['critical_gaps'].split('!')
    )
    change_gaps_ref = create_external_reference(
        SOURCE_WORKBOOKS['change'],
        *EXTERNAL_REFERENCES['change']['critical_gaps'].split('!')
    )
    monitoring_gaps_ref = create_external_reference(
        SOURCE_WORKBOOKS['monitoring'],
        *EXTERNAL_REFERENCES['monitoring']['critical_gaps'].split('!')
    )
    hardening_gaps_ref = create_external_reference(
        SOURCE_WORKBOOKS['hardening'],
        *EXTERNAL_REFERENCES['hardening']['critical_gaps'].split('!')
    )
    
    total_gaps_formula = f'={baseline_gaps_ref}+{change_gaps_ref}+{monitoring_gaps_ref}+{hardening_gaps_ref}'
    
    ws[f'B{row}'] = (
        f'=IF(AND(B{compliance_cell_row}>={COMPLIANCE_TARGETS["overall"]/100},{total_gaps_formula}=0),"Compliant",'
        f'IF(B{compliance_cell_row}>=0.9,"Substantially Compliant",'
        f'IF(B{compliance_cell_row}>=0.8,"At Risk","Non-Compliant")))'
    )
    ws[f'B{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws[f'B{row}'].alignment = ALIGN_CENTER
    ws.merge_cells(f'B{row}:C{row}')
    row += 2
    
    # Section 2: Domain Compliance Breakdown
    ws[f'A{row}'] = 'DOMAIN COMPLIANCE BREAKDOWN'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Table headers
    domain_headers = ['Domain', 'Weight', 'Compliance %', 'Target', 'Status', 'Critical Gaps']
    for idx, header in enumerate(domain_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    row += 1
    
    # Domain rows
    domains = [
        ('Baseline Configuration', DOMAIN_WEIGHTS['baseline_configuration'], baseline_ref, baseline_gaps_ref),
        ('Change Control', DOMAIN_WEIGHTS['change_control'], change_ref, change_gaps_ref),
        ('Configuration Monitoring', DOMAIN_WEIGHTS['configuration_monitoring'], monitoring_ref, monitoring_gaps_ref),
        ('Security Hardening', DOMAIN_WEIGHTS['security_hardening'], hardening_ref, hardening_gaps_ref),
        ('Process Maturity', DOMAIN_WEIGHTS['process_maturity'], process_maturity_formula, '0'),
    ]
    
    domain_start_row = row
    for domain, weight, compliance_ref, gaps_ref in domains:
        ws[f'A{row}'] = domain
        ws[f'B{row}'] = weight
        ws[f'B{row}'].number_format = '0%'
        ws[f'C{row}'] = f'={compliance_ref}'
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = f'≥{COMPLIANCE_TARGETS["domain_minimum"]}%'
        ws[f'D{row}'].alignment = ALIGN_CENTER
        
        # Status
        ws[f'E{row}'] = f'=IF(C{row}>={COMPLIANCE_TARGETS["domain_minimum"]/100},"Green","Red")'
        ws[f'E{row}'].alignment = ALIGN_CENTER
        
        # Critical Gaps
        ws[f'F{row}'] = f'={gaps_ref}' if gaps_ref != '0' else '0'
        ws[f'F{row}'].number_format = '0'
        ws[f'F{row}'].alignment = ALIGN_CENTER
        
        row += 1
    
    row += 1
    
    # Section 3: Key Performance Indicators
    ws[f'A{row}'] = 'KEY PERFORMANCE INDICATORS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    kpi_headers = ['KPI', 'Current', 'Target', 'Status']
    for idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    row += 1
    
    kpis = [
        ('Overall Compliance %', f'=B{compliance_cell_row}', f'≥{COMPLIANCE_TARGETS["overall"]}%', '0.0%'),
        ('Domains Compliant (≥90%)', f'=COUNTIF(E{domain_start_row}:E{domain_start_row+4},"Green")', '5/5', '0'),
        ('Total Critical Gaps', total_gaps_formula, '0', '0'),
        ('Evidence Completeness', '85%', f'≥{COMPLIANCE_TARGETS["evidence_completeness"]}%', '0%'),  # Placeholder
    ]
    
    for kpi_name, current, target, format_str in kpis:
        ws[f'A{row}'] = kpi_name
        ws[f'B{row}'] = current
        ws[f'B{row}'].number_format = format_str
        ws[f'B{row}'].alignment = ALIGN_CENTER
        ws[f'C{row}'] = target
        ws[f'C{row}'].alignment = ALIGN_CENTER
        ws[f'D{row}'] = 'Manual Review'
        ws[f'D{row}'].alignment = ALIGN_CENTER
        row += 1
    
    # Conditional formatting
    # Overall compliance status
    ws.conditional_formatting.add(
        f'B{compliance_cell_row+1}',
        CellIsRule(operator='equal', formula=['"Compliant"'], fill=FILL_GREEN, font=Font(color='FFFFFF', bold=True))
    )
    ws.conditional_formatting.add(
        f'B{compliance_cell_row+1}',
        CellIsRule(operator='equal', formula=['"Non-Compliant"'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )
    
    # Domain status
    ws.conditional_formatting.add(
        f'E{domain_start_row}:E{domain_start_row+4}',
        CellIsRule(operator='equal', formula=['"Green"'], fill=FILL_GREEN)
    )
    ws.conditional_formatting.add(
        f'E{domain_start_row}:E{domain_start_row+4}',
        CellIsRule(operator='equal', formula=['"Red"'], fill=FILL_RED)
    )
    
    # Critical gaps > 0
    ws.conditional_formatting.add(
        f'F{domain_start_row}:F{domain_start_row+4}',
        CellIsRule(operator='greaterThan', formula=['0'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )

# =============================================================================
# SHEET GENERATION FUNCTIONS - PART 2
# =============================================================================

def create_asset_compliance_summary(wb: Workbook) -> None:
    """
    Create the Asset_Compliance_Summary sheet.
    
    This sheet aggregates per-asset compliance across all domains.
    """
    ws = wb.create_sheet("Asset_Compliance_Summary")
    
    # Define headers
    headers = [
        'Asset_ID',
        'Asset_Name',
        'Asset_Type',
        'Asset_Tier',
        'Asset_Owner',
        'Baseline_Documented',
        'Baseline_Current',
        'Change_Compliance',
        'Emergency_Changes_30d',
        'Monitoring_Coverage',
        'Critical_Drift',
        'Hardening_Compliance',
        'High_Risk_Gaps',
        'Overall_Compliance',
        'Status',
        'Critical_Issues',
        'Primary_Gap_Domain',
        'Priority',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Asset_ID
        'B': 30,  # Asset_Name
        'C': 25,  # Asset_Type
        'D': 12,  # Asset_Tier
        'E': 25,  # Asset_Owner
        'F': 18,  # Baseline_Documented
        'G': 16,  # Baseline_Current
        'H': 18,  # Change_Compliance
        'I': 20,  # Emergency_Changes_30d
        'J': 18,  # Monitoring_Coverage
        'K': 14,  # Critical_Drift
        'L': 20,  # Hardening_Compliance
        'M': 16,  # High_Risk_Gaps
        'N': 18,  # Overall_Compliance
        'O': 20,  # Status
        'P': 16,  # Critical_Issues
        'Q': 20,  # Primary_Gap_Domain
        'R': 12,  # Priority
        'S': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add formulas for 100 rows
    # NOTE: These formulas attempt to pull from external workbooks
    # In practice, this requires VLOOKUP across external workbooks which is complex
    # This provides the STRUCTURE - users may need to populate manually or via advanced formulas
    
    for row in range(2, 102):
        # Asset_ID - users enter or pull from baseline
        baseline_ref = create_external_reference(
            SOURCE_WORKBOOKS['baseline'],
            'Asset_Inventory',
            f'A{row}'
        )
        ws[f'A{row}'] = f'={baseline_ref}'
        
        # Asset_Name
        ws[f'B{row}'] = f'={create_external_reference(SOURCE_WORKBOOKS["baseline"], "Asset_Inventory", f"B{row}")}'
        
        # Asset_Type
        ws[f'C{row}'] = f'={create_external_reference(SOURCE_WORKBOOKS["baseline"], "Asset_Inventory", f"C{row}")}'
        
        # Asset_Tier
        ws[f'D{row}'] = f'={create_external_reference(SOURCE_WORKBOOKS["baseline"], "Asset_Inventory", f"D{row}")}'
        
        # Asset_Owner
        ws[f'E{row}'] = f'={create_external_reference(SOURCE_WORKBOOKS["baseline"], "Asset_Inventory", f"E{row}")}'
        
        # Baseline_Documented (Yes/No)
        ws[f'F{row}'] = f'={create_external_reference(SOURCE_WORKBOOKS["baseline"], "Asset_Inventory", f"K{row}")}'
        
        # Baseline_Current (Yes/No)
        ws[f'G{row}'] = f'={create_external_reference(SOURCE_WORKBOOKS["baseline"], "Asset_Inventory", f"L{row}")}'
        
        # Change_Compliance - NOTE: Requires VLOOKUP to match Asset_ID in A.8.9.2
        # Simplified: Users should manually link or use advanced VLOOKUP
        ws[f'H{row}'].value = 'Manual'
        
        # Emergency_Changes_30d - Manual from A.8.9.2
        ws[f'I{row}'].value = 0
        
        # Monitoring_Coverage - Manual from A.8.9.3
        ws[f'J{row}'].value = 'Manual'
        
        # Critical_Drift - Manual from A.8.9.3
        ws[f'K{row}'].value = 0
        
        # Hardening_Compliance - Manual from A.8.9.4
        ws[f'L{row}'].value = 'Manual'
        
        # High_Risk_Gaps - Manual from A.8.9.4
        ws[f'M{row}'].value = 0
        
        # Overall_Compliance calculation
        ws[f'N{row}'] = (
            f'=IF(A{row}="","",('
            f'(IF(AND(F{row}="Yes",G{row}="Yes"),100,IF(OR(F{row}="Yes",G{row}="Yes"),50,0))*0.2)+'
            f'(IF(ISNUMBER(H{row}),H{row},0)*0.25)+'
            f'(IF(AND(J{row}="Yes",K{row}=0),100,IF(J{row}="Yes",50,0))*0.2)+'
            f'(IF(ISNUMBER(L{row}),L{row},0)*0.25)+'
            f'(IF(I{row}<=2,100,100-(I{row}-2)*10)*0.1))/100)'
        )
        ws[f'N{row}'].number_format = '0.0%'
        ws[f'N{row}'].protection = Protection(locked=True)
        
        # Status
        ws[f'O{row}'] = (
            f'=IF(A{row}="","",IF(AND(N{row}>=0.95,P{row}=0),"Fully Compliant",'
            f'IF(N{row}>=0.9,"Compliant",IF(N{row}>=0.8,"Substantially Compliant",'
            f'IF(N{row}>=0.7,"Partially Compliant","Non-Compliant")))))'
        )
        ws[f'O{row}'].protection = Protection(locked=True)
        
        # Critical_Issues count
        ws[f'P{row}'] = (
            f'=IF(A{row}="","",'
            f'(IF(F{row}="No",1,0))+'
            f'(IF(K{row}>0,1,0))+'
            f'(IF(M{row}>0,1,0))+'
            f'(IF(I{row}>5,1,0)))'
        )
        ws[f'P{row}'].number_format = '0'
        ws[f'P{row}'].protection = Protection(locked=True)
        
        # Primary_Gap_Domain - which domain has lowest score
        ws[f'Q{row}'] = (
            f'=IF(A{row}="","","Manual Analysis Required")'
        )
        
        # Priority
        ws[f'R{row}'] = (
            f'=IF(A{row}="","",IF(AND(D{row}="Critical",P{row}>0),"P0",'
            f'IF(P{row}>0,"P1",IF(N{row}<0.9,"P2","P3"))))'
        )
        ws[f'R{row}'].protection = Protection(locked=True)
    
    # Conditional formatting
    # Status colors
    ws.conditional_formatting.add(
        'O2:O101',
        CellIsRule(operator='equal', formula=['"Fully Compliant"'], fill=FILL_DARK_GREEN, font=Font(color='FFFFFF'))
    )
    ws.conditional_formatting.add(
        'O2:O101',
        CellIsRule(operator='equal', formula=['"Compliant"'], fill=FILL_GREEN)
    )
    ws.conditional_formatting.add(
        'O2:O101',
        CellIsRule(operator='equal', formula=['"Substantially Compliant"'], fill=FILL_YELLOW)
    )
    ws.conditional_formatting.add(
        'O2:O101',
        CellIsRule(operator='equal', formula=['"Partially Compliant"'], fill=FILL_ORANGE)
    )
    ws.conditional_formatting.add(
        'O2:O101',
        CellIsRule(operator='equal', formula=['"Non-Compliant"'], fill=FILL_RED, font=Font(color='FFFFFF'))
    )
    
    # Critical Issues > 0
    ws.conditional_formatting.add(
        'P2:P101',
        CellIsRule(operator='greaterThan', formula=['0'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )
    
    # Priority P0/P1
    ws.conditional_formatting.add(
        'R2:R101',
        CellIsRule(operator='equal', formula=['"P0"'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )
    ws.conditional_formatting.add(
        'R2:R101',
        CellIsRule(operator='equal', formula=['"P1"'], fill=FILL_ORANGE)
    )
    
    # Unlock manual data entry cells
    unlock_cell_range(ws, 'H2:M101')  # Manual fields
    unlock_cell_range(ws, 'S2:S101')  # Notes


def create_gap_prioritization(wb: Workbook) -> None:
    """
    Create the Gap_Prioritization sheet.
    
    This sheet consolidates gaps from all 4 domains with risk-based prioritization.
    """
    ws = wb.create_sheet("Gap_Prioritization")
    
    # Define headers
    headers = [
        'Gap_ID',
        'Priority_Rank',
        'Source_Domain',
        'Asset_ID',
        'Asset_Name',
        'Asset_Tier',
        'Gap_Category',
        'Gap_Description',
        'Gap_Risk_Rating',
        'Impact',
        'Exploitation_Likelihood',
        'Risk_Score',
        'Owner',
        'Strategy',
        'Target_Date',
        'Days_Until',
        'Status',
        'Quick_Win',
        'Cross_Domain',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Gap_ID
        'B': 12,  # Priority_Rank
        'C': 20,  # Source_Domain
        'D': 12,  # Asset_ID
        'E': 25,  # Asset_Name
        'F': 12,  # Asset_Tier
        'G': 20,  # Gap_Category
        'H': 50,  # Gap_Description
        'I': 16,  # Gap_Risk_Rating
        'J': 40,  # Impact
        'K': 20,  # Exploitation_Likelihood
        'L': 12,  # Risk_Score
        'M': 25,  # Owner
        'N': 30,  # Strategy
        'O': 15,  # Target_Date
        'P': 12,  # Days_Until
        'Q': 18,  # Status
        'R': 12,  # Quick_Win
        'S': 14,  # Cross_Domain
        'T': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    domain_list = [
        'Baseline Configuration',
        'Change Control',
        'Configuration Monitoring',
        'Security Hardening',
        'Process Maturity'
    ]
    
    gap_categories = [
        'Baseline Not Documented',
        'Baseline Out of Date',
        'Change Control Non-Compliance',
        'Unauthorised Change',
        'Configuration Drift',
        'Monitoring Gap',
        'Hardening Control Missing',
        'Exception Without Justification',
        'Process Deficiency'
    ]
    
    risk_ratings = ['Critical', 'High', 'Medium', 'Low']
    likelihood = ['Very High', 'High', 'Medium', 'Low', 'Very Low']
    status_list = ['Identified', 'Planning', 'In Progress', 'Blocked', 'Completed', 'Closed']
    yes_no = ['Yes', 'No']
    
    # Apply data validation
    for dv_range, dv_list in [
        ('C2:C151', domain_list),
        ('G2:G151', gap_categories),
        ('I2:I151', risk_ratings),
        ('K2:K151', likelihood),
        ('Q2:Q151', status_list),
        ('R2:R151', yes_no),
        ('S2:S151', yes_no),
    ]:
        dv = DataValidation(type="list", formula1=f'"{",".join(dv_list)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(dv_range)
    
    # Add formulas for 150 rows
    for row in range(2, 152):
        # Gap_ID auto-generated
        ws[f'A{row}'] = f'=IF(C{row}<>"","GAP-"&TEXT(ROW()-1,"000"),"")'
        ws[f'A{row}'].protection = Protection(locked=True)
        
        # Priority_Rank - manual sorting by Risk_Score
        ws[f'B{row}'] = f'=IF(L{row}<>"",RANK(L{row},L:L,0),"")'
        ws[f'B{row}'].number_format = '0'
        ws[f'B{row}'].protection = Protection(locked=True)
        
        # Risk_Score calculation
        # Weights: Asset Tier (10), Risk Rating (8), Likelihood (6), Days Overdue (2)
        ws[f'L{row}'] = (
            f'=IF(C{row}="","",('
            f'(IF(F{row}="Critical",5,IF(F{row}="High",4,IF(F{row}="Medium",3,2)))*10)+'
            f'(IF(I{row}="Critical",5,IF(I{row}="High",4,IF(I{row}="Medium",3,2)))*8)+'
            f'(IF(K{row}="Very High",5,IF(K{row}="High",4,IF(K{row}="Medium",3,IF(K{row}="Low",2,1))))*6)+'
            f'(IF(P{row}<0,ABS(P{row})/7*2,0))))'
        )
        ws[f'L{row}'].number_format = '0.0'
        ws[f'L{row}'].protection = Protection(locked=True)
        
        # Days_Until calculation
        ws[f'P{row}'] = f'=IF(O{row}<>"",O{row}-TODAY(),"")'
        ws[f'P{row}'].number_format = '0'
        ws[f'P{row}'].protection = Protection(locked=True)
        
        # Quick_Win formula
        ws[f'R{row}'] = (
            f'=IF(AND(OR(I{row}="Medium",I{row}="High",I{row}="Critical"),'
            f'OR(N{row}="Configuration Change",N{row}="Policy Update",P{row}<=7)),"Yes","No")'
        )
        ws[f'R{row}'].protection = Protection(locked=True)
    
    # Conditional formatting
    # Priority ranks 1-10: Red
    ws.conditional_formatting.add(
        'B2:B151',
        CellIsRule(operator='between', formula=['1', '10'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )
    
    # Priority ranks 11-30: Orange
    ws.conditional_formatting.add(
        'B2:B151',
        CellIsRule(operator='between', formula=['11', '30'], fill=FILL_ORANGE)
    )
    
    # Quick Win: Green border
    ws.conditional_formatting.add(
        'A2:T151',
        FormulaRule(
            formula=['$R2="Yes"'],
            border=Border(
                left=Side(style='medium', color='70AD47'),
                right=Side(style='medium', color='70AD47'),
                top=Side(style='medium', color='70AD47'),
                bottom=Side(style='medium', color='70AD47')
            )
        )
    )
    
    # Days overdue (negative): Red text
    ws.conditional_formatting.add(
        'P2:P151',
        CellIsRule(operator='lessThan', formula=['0'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )
    
    # Status = Blocked: Gray
    ws.conditional_formatting.add(
        'Q2:Q151',
        CellIsRule(operator='equal', formula=['"Blocked"'], fill=FILL_GRAY)
    )
    
    # Cross-Domain impact: Blue
    ws.conditional_formatting.add(
        'S2:S151',
        CellIsRule(operator='equal', formula=['"Yes"'], fill=PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid'))
    )
    
    # Unlock data entry cells (all except formulas)
    unlock_cell_range(ws, 'C2:K151')  # Source through Likelihood
    unlock_cell_range(ws, 'M2:O151')  # Owner through Target_Date
    unlock_cell_range(ws, 'Q2:Q151')  # Status
    unlock_cell_range(ws, 'S2:T151')  # Cross_Domain, Notes

# =============================================================================
# SHEET GENERATION FUNCTIONS - PART 3
# =============================================================================

def create_trend_analysis(wb: Workbook) -> None:
    """
    Create the Trend_Analysis sheet.
    
    This sheet tracks compliance trajectory over time.
    """
    ws = wb.create_sheet("Trend_Analysis")
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    
    row = 1
    
    # Title
    ws[f'A{row}'] = 'Configuration Management Compliance Trend Analysis'
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells(f'A{row}:H{row}')
    row += 2
    
    # Section 1: Assessment History
    ws[f'A{row}'] = 'ASSESSMENT HISTORY'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Headers for history table
    history_headers = [
        'Cycle',
        'Date',
        'Baseline %',
        'Change %',
        'Monitor %',
        'Harden %',
        'Process %',
        'Overall %'
    ]
    
    for idx, header in enumerate(history_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    row += 1
    
    # History rows (20 cycles capacity = 5 years quarterly)
    history_start_row = row
    for cycle in range(1, 21):
        ws[f'A{row}'] = cycle
        ws[f'A{row}'].alignment = ALIGN_CENTER
        
        # Date
        ws[f'B{row}'].number_format = 'DD.MM.YYYY'
        
        # Percentages
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{row}'].number_format = '0.0%'
        
        row += 1
    
    row += 1
    
    # Section 2: Trend Indicators
    ws[f'A{row}'] = 'TREND INDICATORS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    trend_headers = ['Domain', 'Current %', 'Previous %', 'Change', 'Trend', 'Status']
    for idx, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    row += 1
    
    # Trend rows
    domains = [
        'Baseline Configuration',
        'Change Control',
        'Configuration Monitoring',
        'Security Hardening',
        'Process Maturity',
        'OVERALL COMPLIANCE'
    ]
    
    trend_start_row = row
    for idx, domain in enumerate(domains):
        ws[f'A{row}'] = domain
        if domain == 'OVERALL COMPLIANCE':
            ws[f'A{row}'].font = FONT_BOLD
        
        # Current % - pull from most recent history entry
        col_letter = get_column_letter(3 + idx) if idx < 6 else 'H'
        ws[f'B{row}'] = f'=INDEX({col_letter}:{col_letter},MATCH(MAX(B:B),B:B,0))'
        ws[f'B{row}'].number_format = '0.0%'
        
        # Previous % - pull from second most recent
        ws[f'C{row}'] = f'=INDEX({col_letter}:{col_letter},MATCH(MAX(B:B),B:B,0)-1)'
        ws[f'C{row}'].number_format = '0.0%'
        
        # Change
        ws[f'D{row}'] = f'=IF(C{row}<>"",B{row}-C{row},"")'
        ws[f'D{row}'].number_format = '+0.0%;-0.0%;0.0%'
        
        # Trend arrow
        ws[f'E{row}'] = f'=IF(D{row}="","",IF(D{row}>0.02,"↑ Improving",IF(D{row}<-0.02,"↓ Declining","→ Stable")))'
        
        # Status
        ws[f'F{row}'] = f'=IF(B{row}="","",IF(AND(B{row}>=0.9,OR(E{row}="↑ Improving",E{row}="→ Stable")),"Green",IF(OR(B{row}<0.8,D{row}<-0.05),"Red","Yellow")))'
        
        row += 1
    
    row += 1
    
    # Section 3: Gap Trend
    ws[f'A{row}'] = 'GAP TREND ANALYSIS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    gap_headers = ['Cycle', 'Total Gaps', 'Critical', 'High', 'Closed', 'Opened', 'Net Change']
    for idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    row += 1
    
    # Gap trend rows (10 cycles)
    for cycle in range(1, 11):
        ws[f'A{row}'] = cycle
        ws[f'A{row}'].alignment = ALIGN_CENTER
        
        # Net Change formula (if not first row)
        if cycle > 1:
            ws[f'G{row}'] = f'=IF(AND(E{row}<>"",F{row}<>""),E{row}-F{row},"")'
            ws[f'G{row}'].number_format = '+0;-0;0'
        
        row += 1
    
    row += 1
    
    # Section 4: First Assessment Guidance
    ws[f'A{row}'] = 'FIRST ASSESSMENT GUIDANCE'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_LIGHT_YELLOW
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    guidance = [
        'This is your baseline assessment. Trend analysis requires 3+ assessment cycles.',
        '',
        'Next Steps:',
        '1. Schedule next assessment (recommend quarterly)',
        '2. After 2nd assessment: Compare Current vs. Baseline',
        '3. After 3rd assessment: Begin trend analysis',
        '4. After 4th assessment: Predictive analysis becomes available',
        '',
        'Maintain consistency in assessment methodology across cycles.',
    ]
    
    for line in guidance:
        ws[f'A{row}'] = line
        ws[f'A{row}'].alignment = ALIGN_LEFT
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
    
    row += 1
    
    ws[f'A{row}'] = 'Current Assessment Baseline:'
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    # Pull current compliance from dashboard
    ws[f'A{row}'] = 'Overall Compliance:'
    ws[f'B{row}'] = '=Overall_Compliance_Dashboard!B6'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = Font(name='Calibri', size=12, bold=True)
    row += 1
    
    ws[f'A{row}'] = 'Total Critical Gaps:'
    ws[f'B{row}'] = '=SUM(Overall_Compliance_Dashboard!F12:F16)'
    ws[f'B{row}'].number_format = '0'
    row += 1
    
    ws[f'A{row}'] = 'Target Next Cycle:'
    ws[f'B{row}'] = 'Improve by ≥5 percentage points OR close ≥50% of gaps'
    ws.merge_cells(f'B{row}:H{row}')
    
    # Conditional formatting
    # Trend status
    ws.conditional_formatting.add(
        f'F{trend_start_row}:F{trend_start_row+5}',
        CellIsRule(operator='equal', formula=['"Green"'], fill=FILL_GREEN)
    )
    ws.conditional_formatting.add(
        f'F{trend_start_row}:F{trend_start_row+5}',
        CellIsRule(operator='equal', formula=['"Red"'], fill=FILL_RED)
    )
    ws.conditional_formatting.add(
        f'F{trend_start_row}:F{trend_start_row+5}',
        CellIsRule(operator='equal', formula=['"Yellow"'], fill=FILL_YELLOW)
    )
    
    # Unlock history data entry
    unlock_cell_range(ws, f'B{history_start_row}:H{history_start_row+19}')


def create_evidence_register(wb: Workbook) -> None:
    """
    Create the Evidence_Register sheet.
    
    Consolidated evidence index linking to domain-specific evidence.
    """
    ws = wb.create_sheet("Evidence_Register")
    
    # Define headers
    headers = [
        'Evidence_ID',
        'Source_Domain',
        'Source_Document',
        'Source_Evidence_ID',
        'Evidence_Type',
        'Evidence_Description',
        'Related_Control_Area',
        'Collection_Date',
        'Evidence_Location',
        'Evidence_Status',
        'Verification_Status',
        'Audit_Reference',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 14,  # Evidence_ID
        'B': 22,  # Source_Domain
        'C': 20,  # Source_Document
        'D': 18,  # Source_Evidence_ID
        'E': 25,  # Evidence_Type
        'F': 50,  # Evidence_Description
        'G': 25,  # Related_Control_Area
        'H': 15,  # Collection_Date
        'I': 40,  # Evidence_Location
        'J': 15,  # Evidence_Status
        'K': 18,  # Verification_Status
        'L': 20,  # Audit_Reference
        'M': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    domain_list = [
        'Baseline Configuration',
        'Change Control',
        'Configuration Monitoring',
        'Security Hardening',
        'Dashboard/Overall'
    ]
    
    evidence_types = [
        'Baseline Documentation',
        'Configuration Export',
        'Change Record',
        'Approval Documentation',
        'Monitoring Report',
        'Alert Configuration',
        'Drift Detection Log',
        'Hardening Assessment Report',
        'Exception Approval',
        'Risk Assessment',
        'Process Documentation',
        'Training Records',
        'Audit Report',
        'Management Review',
    ]
    
    status_list = ['Active', 'Expired', 'Superseded']
    verification_list = ['Verified', 'Pending', 'Failed']
    
    # Apply data validation
    for dv_range, dv_list in [
        ('B2:B101', domain_list),
        ('E2:E101', evidence_types),
        ('J2:J101', status_list),
        ('K2:K101', verification_list),
    ]:
        dv = DataValidation(type="list", formula1=f'"{",".join(dv_list)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(dv_range)
    
    # Add formulas for 100 rows
    for row in range(2, 102):
        # Evidence_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"","EVD-"&TEXT(ROW()-1,"000"),"")'
        ws[f'A{row}'].protection = Protection(locked=True)
        
        # Collection_Date formatting
        ws[f'H{row}'].number_format = 'DD.MM.YYYY'
        
        # Default values
        ws[f'J{row}'].value = 'Active'
        ws[f'K{row}'].value = 'Pending'
    
    # Conditional formatting
    # Status
    ws.conditional_formatting.add(
        'J2:J101',
        CellIsRule(operator='equal', formula=['"Expired"'], fill=FILL_RED)
    )
    ws.conditional_formatting.add(
        'J2:J101',
        CellIsRule(operator='equal', formula=['"Superseded"'], fill=FILL_GRAY)
    )
    
    # Verification Status
    ws.conditional_formatting.add(
        'K2:K101',
        CellIsRule(operator='equal', formula=['"Verified"'], fill=FILL_GREEN)
    )
    ws.conditional_formatting.add(
        'K2:K101',
        CellIsRule(operator='equal', formula=['"Failed"'], fill=FILL_RED)
    )
    ws.conditional_formatting.add(
        'K2:K101',
        CellIsRule(operator='equal', formula=['"Pending"'], fill=FILL_YELLOW)
    )
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'B2:I101')  # Source through Location
    unlock_cell_range(ws, 'J2:M101')  # Status through Notes


def create_approval_sign_off(wb: Workbook) -> None:
    """Create the Approval_Sign_Off sheet."""
    ws = wb.create_sheet("Approval_Sign_Off")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    
    row = 1
    
    # Title
    ws[f'A{row}'] = 'Configuration Management Compliance'
    ws[f'A{row}'].font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Approval & Sign-Off'
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells(f'A{row}:D{row}')
    row += 2
    
    # Section 1: Dashboard Summary
    ws[f'A{row}'] = 'DASHBOARD SUMMARY'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    summary_items = [
        ('Assessment Period', CONFIG['dashboard_date']),
        ('Overall Compliance', '=Overall_Compliance_Dashboard!B6'),
        ('Compliance Status', '=Overall_Compliance_Dashboard!B7'),
        ('Total Critical Gaps', '=SUM(Overall_Compliance_Dashboard!F12:F16)'),
        ('Domains Compliant', '=Overall_Compliance_Dashboard!B19&"/5"'),
    ]
    
    for label, value in summary_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = FONT_BOLD
        if isinstance(value, str) and value.startswith('='):
            ws[f'B{row}'] = value
            if 'Compliance' in label and 'Status' not in label:
                ws[f'B{row}'].number_format = '0.0%'
        else:
            ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # Section 2: Executive Summary
    ws[f'A{row}'] = 'EXECUTIVE SUMMARY'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = 'Summary:'
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    summary_text = (
        "[Provide narrative summary including:\n"
        "\u2022 Overall compliance status and trajectory\n"
        "\u2022 Key achievements since last review\n"
        "\u2022 Critical gaps requiring attention\n"
        "\u2022 Resource requirements\n"
        "\u2022 Recommended actions]"
    )
    ws[f'A{row}'] = summary_text
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:D{row}')
    ws.row_dimensions[row].height = 100
    row += 2
    
    # Section 3: Three-Tier Approval
    ws[f'A{row}'] = 'APPROVAL SIGN-OFF'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Tier 1
    ws[f'A{row}'] = 'Tier 1 - Operational Review'
    ws[f'A{row}'].font = FONT_BOLD
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    tier_headers = ['Role', 'Name', 'Date', 'Comments']
    for idx, header in enumerate(tier_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
    row += 1
    
    tier1_roles = [
        ('Dashboard Owner', CONFIG['dashboard_owner']),
        ('ISMS Manager', '[Name]'),
    ]
    
    for role, name in tier1_roles:
        ws[f'A{row}'] = role
        ws[f'B{row}'] = name
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'
        ws[f'D{row}'] = ''
        row += 1
    
    row += 1
    
    # Tier 2
    ws[f'A{row}'] = 'Tier 2 - Management Review'
    ws[f'A{row}'].font = FONT_BOLD
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for idx, header in enumerate(tier_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
    row += 1
    
    tier2_roles = [
        ('IT Manager', '[Name]'),
        ('Security Manager', '[Name]'),
    ]
    
    for role, name in tier2_roles:
        ws[f'A{row}'] = role
        ws[f'B{row}'] = name
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'
        ws[f'D{row}'] = ''
        row += 1
    
    row += 1
    
    # Tier 3
    ws[f'A{row}'] = 'Tier 3 - Executive Acceptance'
    ws[f'A{row}'].font = FONT_BOLD
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for idx, header in enumerate(tier_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
    row += 1
    
    ws[f'A{row}'] = 'CISO / CIO'
    ws[f'B{row}'] = '[Name]'
    ws[f'C{row}'].number_format = 'DD.MM.YYYY'
    ws[f'D{row}'] = ''
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'A16:D16')  # Executive Summary
    unlock_cell_range(ws, 'B22:D23')  # Tier 1 approvals
    unlock_cell_range(ws, 'B27:D28')  # Tier 2 approvals
    unlock_cell_range(ws, 'B32:D32')  # Tier 3 approval

# =============================================================================
# MAIN ORCHESTRATION AND EXECUTION
# =============================================================================

def create_all_sheets(wb: Workbook) -> None:
    """Create all sheets with proper implementations."""
    
    logger.info("  [1/8] Instructions...")
    create_instructions_sheet(wb)
    
    logger.info("  [2/8] Workbook_Integration_Settings...")
    create_workbook_integration_settings(wb)
    
    logger.info("  [3/8] Overall_Compliance_Dashboard...")
    create_overall_compliance_dashboard(wb)
    
    logger.info("  [4/8] Asset_Compliance_Summary...")
    create_asset_compliance_summary(wb)
    
    logger.info("  [5/8] Gap_Prioritization...")
    create_gap_prioritization(wb)
    
    logger.info("  [6/8] Trend_Analysis...")
    create_trend_analysis(wb)
    
    logger.info("  [7/8] Evidence_Register...")
    create_evidence_register(wb)
    
    logger.info("  [8/8] Approval_Sign_Off...")
    create_approval_sign_off(wb)


def main():
    """Main execution function."""
    logger.info("=" * 80)
    logger.info("ISMS Control A.8.9.5 - Compliance Dashboard")
    logger.info("Workbook Generation Script")
    logger.info("=" * 80)
    logger.info("")
    
    logger.info(f"Organisation: {CONFIG['organization_name']}")
    logger.info(f"Dashboard Date: {CONFIG['dashboard_date']}")
    logger.info(f"Dashboard Owner: {CONFIG['dashboard_owner']}")
    logger.info(f"Output File: {CONFIG['output_filename']}")
    logger.info("")
    
    logger.info("Creating workbook...")
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    logger.info("Generating sheets:")
    create_all_sheets(wb)
    
    logger.info("")
    logger.info("Saving workbook...")
    wb.save(CONFIG['output_filename'])
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("✓ Workbook generated successfully!")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Output: {CONFIG['output_filename']}")
    logger.info("")
    logger.info("Sheet Summary:")
    logger.info("-" * 40)
    for i, sheet in enumerate(wb.sheetnames, 1):
        logger.info(f"  {i}. {sheet}")
    logger.info("")
    logger.info("Key Capacities:")
    logger.info("-" * 40)
    logger.info("  \u2022 Asset Compliance Summary: 100 assets")
    logger.info("  \u2022 Gap Prioritization: 150 gaps")
    logger.info("  \u2022 Trend Analysis: 20 assessment cycles")
    logger.info("  \u2022 Evidence Register: 100 evidence items")
    logger.info("")
    logger.info("CRITICAL NEXT STEPS:")
    logger.info("=" * 80)
    logger.info("1. Normalize source workbooks:")
    logger.info("   python3 normalize_assessment_files_a89.py")
    logger.info("")
    logger.info("2. Place this dashboard in SAME DIRECTORY as normalized files:")
    logger.info("   - ISMS-IMP-A.8.9.1.xlsx")
    logger.info("   - ISMS-IMP-A.8.9.2.xlsx")
    logger.info("   - ISMS-IMP-A.8.9.3.xlsx")
    logger.info("   - ISMS-IMP-A.8.9.4.xlsx")
    logger.info("")
    logger.info("3. Open dashboard and click 'Update Links' when Excel prompts")
    logger.info("")
    logger.info("4. Verify Workbook_Integration_Settings shows all 4 workbooks linked")
    logger.info("")
    logger.info("5. Review Overall_Compliance_Dashboard for compliance status")
    logger.info("")
    logger.info("6. Manual data population required for:")
    logger.info("   \u2022 Asset_Compliance_Summary: Cross-domain lookups (H-M columns)")
    logger.info("   \u2022 Gap_Prioritization: Aggregate gaps from all sources")
    logger.info("   \u2022 Trend_Analysis: Historical data after each cycle")
    logger.info("   \u2022 Evidence_Register: Consolidate from domain registers")
    logger.info("")
    logger.info("=" * 80)
    logger.info("")
    logger.info("IMPORTANT NOTES:")
    logger.info("\u2022 External workbook links REQUIRE source files in same directory")
    logger.info("\u2022 Source files MUST have normalized names (no dates)")
    logger.error("\u2022 Dashboard will show #REF! errors if links are broken")
    logger.info("\u2022 Use Excel Data → Edit Links to troubleshoot broken links")
    logger.info("\u2022 Some sheets require manual population (see above)")
    logger.info("")
    logger.info("WORKBOOK STRUCTURE:")
    logger.info("-" * 40)
    logger.info("Instructions:")
    logger.info("  → Comprehensive setup and usage guidance")
    logger.info("")
    logger.info("Workbook_Integration_Settings:")
    logger.info("  → Validates external workbook links")
    logger.info("  → Shows link health and troubleshooting")
    logger.info("")
    logger.info("Overall_Compliance_Dashboard:")
    logger.info("  → Executive summary with weighted compliance scoring")
    logger.info("  → Domain breakdown and KPIs")
    logger.info("  → Primary reporting view")
    logger.info("")
    logger.info("Asset_Compliance_Summary:")
    logger.info("  → Per-asset view across all 4 domains")
    logger.info("  → Pulls Asset_ID from A.8.9.1 automatically")
    logger.info("  → Requires manual population of cross-domain data")
    logger.info("")
    logger.info("Gap_Prioritization:")
    logger.info("  → Consolidated gap list with risk-based ranking")
    logger.info("  → Auto-calculates Risk_Score and Priority_Rank")
    logger.info("  → Identifies Quick Wins and Cross-Domain impacts")
    logger.info("")
    logger.info("Trend_Analysis:")
    logger.info("  → Historical compliance tracking")
    logger.info("  → Requires 3+ cycles for meaningful trends")
    logger.info("  → Manual data entry after each assessment cycle")
    logger.info("")
    logger.info("Evidence_Register:")
    logger.info("  → Consolidated evidence index")
    logger.info("  → Links to domain-specific evidence")
    logger.info("  → Auto-generates Evidence_ID")
    logger.info("")
    logger.info("Approval_Sign_Off:")
    logger.info("  → Three-tier approval process")
    logger.info("  → Executive summary and risk assessment")
    logger.info("  → Formal governance documentation")
    logger.info("")
    logger.info("=" * 80)
    logger.info("")
    logger.info("CUSTOMIZATION REMINDERS:")
    logger.info("\u2022 Domain weights: Review DOMAIN_WEIGHTS in CONFIG section")
    logger.info("\u2022 External references: Verify EXTERNAL_REFERENCES match your workbooks")
    logger.info("\u2022 Compliance targets: Adjust COMPLIANCE_TARGETS as needed")
    logger.info("\u2022 Cell references: Update if source workbook structures differ")
    logger.info("")
    logger.info("=" * 80)
    logger.info("")
    logger.info("PROFESSIONAL ISMS IMPLEMENTATION")
    logger.info("Evidence > Theater - Systems Engineering ISMS")
    logger.info("=" * 80)
    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
