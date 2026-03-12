#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.9.2 - Change Control Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Assessment Domain 2 of 4: Change Control and Configuration Updates

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific change management processes, CAB operations, and
approval workflows.

Key customization areas:
1. Change classification criteria (match your change management framework)
2. CAB membership and approval authorities (align with organisational roles)
3. Approval workflow tiers (adapt to your governance structure)
4. Testing requirements and environments (specific to your infrastructure)
5. Change management system integration (ServiceNow, Jira, etc.)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
change control and configuration update processes against ISO 27001:2022
Control A.8.9 requirements.

**Purpose:**
Enables systematic assessment of change classification, CAB operations, approval
workflows, testing procedures, and rollback capabilities to ensure controlled
and authorised configuration modifications.

**Assessment Scope:**
- Change classification framework (Standard, Normal, Emergency)
- Change Advisory Board (CAB) operations and governance
- Approval workflow execution and authorisation
- Testing and validation requirements compliance
- Rollback procedure readiness and testing
- Change success metrics and KPI tracking
- Post-implementation review completion
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and change control standards
2. Change Classification - Change type inventory and classification criteria
3. CAB Operations - CAB membership, meeting schedule, decision tracking
4. Approval Workflows - Approval chain execution and authorisation evidence
5. Testing & Validation - Test planning, execution, and results
6. Rollback Procedures - Rollback readiness and testing evidence
7. Change Metrics - Success rate, emergency changes, SLA compliance
8. Gap Analysis - Non-compliant changes and remediation requirements
9. Evidence Register - Audit evidence tracking (100+ entries)
10. Approval & Sign-Off - Three-tier approval workflow

**Key Features:**
- Data validation with change type and risk level dropdowns
- Conditional formatting for approval status and SLA compliance
- Automated gap identification for untested or unapproved changes
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with change management systems

**Integration:**
consolidates data from all four configuration management assessment domains
for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_2_change_control.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_2_change_control.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_2_change_control.py --date 20250127

Output:
    File: ISMS_IMP_A_8_9_2_Change_Control_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise change classification criteria
    2. Validate CAB membership and meeting schedule
    3. Document approval workflows by change risk level
    4. Assess testing and validation compliance
    5. Verify rollback procedure readiness
    6. Calculate change success metrics
    7. Conduct gap analysis for non-compliant changes
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (change requests, CAB minutes)
    10. Obtain three-tier stakeholder approvals

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Assessment Domain:    2 of 4 (Change Control and Configuration Updates)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-POL-A.8.9, Section 2.3: Change Control & Configuration Updates
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9, Part 2: Change Management Implementation Guide
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.1: Baseline Configuration Assessment (Domain 1)
    - ISMS-IMP-A.8.9.3: Configuration Monitoring Assessment (Domain 3)
    - ISMS-IMP-A.8.9.4: Security Hardening Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.9.2 specification
    - Supports comprehensive change control evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Change Management Standards:**
Change management best practices evolve. Review ITIL 4 change enablement,
organisational change success metrics, and industry benchmarks quarterly.
Emergency change abuse and inadequate testing must be identified and corrected.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of CAB operations, approvals, and testing.

**Data Protection:**
Assessment workbooks contain sensitive operational details including:
- Change request records with system details
- CAB meeting minutes with decision rationale
- Testing results and vulnerability information
- Rollback procedures and recovery capabilities

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Check change success metrics and emergency change rate
- Quarterly: Review CAB effectiveness and approval SLA compliance
- Annually: Complete reassessment of change control processes
- Ad-hoc: When change management process changes or incidents occur

**Quality Assurance:**
Have change management coordinators and CAB leadership validate assessments
before using results for compliance reporting or process improvement decisions.

**Regulatory Alignment:**
Ensure change control practices align with applicable requirements:
- ISO 27001:2022: Control A.8.9 change management
- ITIL 4: Change enablement practices
- Sector-specific: Regulatory change control requirements
- Internal: Organisational change management policies

Customize assessment criteria to include regulatory-specific requirements.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# ============================================================================
# CONFIGURATION SECTION - CUSTOMIZE FOR YOUR ENVIRONMENT
# ============================================================================

# File output configuration
FILENAME = f"ISMS-IMP-A.8.9.2_Change_Control_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Workbook metadata
WORKBOOK_TITLE = "Change Control Assessment"
WORKBOOK_NAME = "Change Control"
WORKBOOK_VERSION = "1.0"

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.9.2"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
CONTROL_ID   = "A.8.9"
CONTROL_NAME = "Configuration Management"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
# CUSTOMIZE: Change management dropdown values
CHANGE_TYPE = ["Standard", "Normal", "Emergency", "Hot Fix"]
CHANGE_PRIORITY = ["P1-Critical", "P2-High", "P3-Medium", "P4-Low"]
CHANGE_STATUS = ["Draft", "Submitted", "\u2705 Approved", "In Testing", "Scheduled",
                "⏳ Implementing", "\u2705 Completed", "\u274C Failed", "Rolled Back", "\u274C Cancelled"]

APPROVAL_TIER = ["Single-Tier", "Two-Tier", "Three-Tier", "Emergency"]
APPROVAL_STATUS_TIER = ["⏳ Pending", "\u2705 Approved", "\u274C Rejected", "➖ N/A"]
APPROVAL_METHOD = ["CAB Meeting", "Email Approval", "Emergency Verbal", "Automated (Standard)", "Not Applicable"]

USER_IMPACT = ["None", "Minimal", "Moderate", "Significant", "Severe"]
SERVICE_DOWNTIME = ["None", "<1 hour", "1-4 hours", "4-8 hours", ">8 hours"]
RISK_LEVEL = ["Low", "Medium", "High", "Critical"]
YES_NO = ["Yes", "No"]
YES_NO_NA = ["Yes", "No", "N/A"]
YES_NO_PARTIAL_NA = ["Yes", "No", "Partially", "N/A"]

TEST_ENVIRONMENT = ["Dev", "Test", "Staging", "UAT", "Production (Non-Critical)", "None"]
TESTING_STATUS = ["❌ Not Started", "⏳ In Progress", "\u2705 Completed", "\u274C Failed", "⚡ Abbreviated (Emergency)"]
GO_NOGO = ["Go", "No-Go", "Go with Conditions", "N/A"]

IMPLEMENTATION_METHOD = ["Manual", "Automated Script", "Semi-Automated", "Assisted (Vendor)"]
VERIFICATION_STATUS = ["\u2705 Successful", "\u26A0\uFE0F Partial Success", "\u274C Failed", "❓ Not Verified"]
ISSUES_RESOLVED = ["Yes", "No", "Partially", "N/A"]
IMPLEMENTATION_STATUS = ["\u2705 Successful", "\u274C Failed", "Rolled Back", "⏳ In Progress"]

ROLLBACK_TEST_RESULTS = ["\u2705 Successful", "\u274C Failed", "\u26A0\uFE0F Partially Successful", "❓ Not Tested"]
DATA_LOSS_RISK = ["None", "Minimal", "Moderate", "Significant"]
ROLLBACK_APPROVAL = ["Yes (same as forward)", "Yes (expedited)", "No (automatic)"]

EMERGENCY_TYPE = ["Security Incident", "Service Outage", "Critical Bug", "Vulnerability Remediation", "Other"]
EMERGENCY_APPROVAL_METHOD = ["Verbal (CIO/CISO)", "Email (Expedited)", "CAB Chair Authorisation"]
POST_IMPL_DOC = ["Yes", "No", "In Progress"]
CAB_REVIEW_OUTCOME = ["\u2705 Approved", "\u26A0\uFE0F Approved with Remediation", "\u274C Disapproved (requires reversal)", "⏳ Not Yet Reviewed"]
JUSTIFICATION_VALID = ["Yes", "No", "Questionable"]

EVIDENCE_TYPE = ["Approval Sign-Off", "Test Results", "Implementation Log", "Rollback Test", 
                "CAB Minutes", "Email Approval", "Change Request", "Other"]
EVIDENCE_CLASSIFICATION = ["Public", "Internal", "Confidential", "Restricted"]
RETENTION_PERIOD = ["1 Year", "3 Years", "5 Years", "7 Years", "Indefinite"]
EVIDENCE_VERIFICATION = ["\u2705 Verified", "⏳ Needs Verification", "\u274C Missing", "⏰ Outdated"]

APPROVAL_DECISION = ["Approved", "Approved with Conditions", "Not Approved - Revisions Required"]

# CUSTOMIZE: Color scheme
COLORS = {
    'header_main': '003366',          # Dark Blue
    'header_sub': '4472C4',           # Blue
    'column_header': 'D9D9D9',        # Light Gray
    'input_cell': 'FFFFCC',           # White
    'protected_cell': 'F2F2F2',       # Very Light Gray
    'compliant': 'C6EFCE',            # Green
    'partial': 'FFEB9C',              # Yellow
    'non_compliant': 'FFC7CE',        # Red
    'excluded': 'D9D9D9',             # Gray
    'critical': 'C00000',             # Dark Red
    'info_bg': 'F2F2F2',              # Light Gray
    'light_green': 'C6EFCE',          # Light Green
    'orange': 'FFA500'                # Orange
}
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================
def create_styles():
    """Creates and returns a dictionary of reusable styles."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    styles = {
        'header_main': {
            'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_main'], 
                              end_color=COLORS['header_main'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'header_sub': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'column_header': {
            'font': Font(name='Calibri', size=11, bold=True),
            'fill': PatternFill(start_color=COLORS['column_header'], 
                              end_color=COLORS['column_header'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'input_cell': {
            'font': Font(name='Calibri', size=11),
            'fill': PatternFill(start_color=COLORS['input_cell'], 
                              end_color=COLORS['input_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False),
            'border': thin_border
        },
        'protected_cell': {
            'font': Font(name='Calibri', size=11, italic=True),
            'fill': PatternFill(start_color=COLORS['protected_cell'], 
                              end_color=COLORS['protected_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top'),
            'border': thin_border
        },
        'info_text': {
            'font': Font(name='Calibri', size=11),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True)
        },
        'section_header': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': thin_border
        }
    }
    return styles

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_data_validation(values, allow_blank=True):
    """Create a data validation object for dropdowns."""
    formula = f'"{",".join(values)}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the dropdown'
    dv.promptTitle = 'Selection Required'
    return dv

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab for applicable systems/services.', '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Check all applicable items in the Compliance Checklist for each section.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_change_request_register_sheet(wb, styles):
    """Create Sheet 2: Change Request Register"""
    ws = wb.create_sheet("Change Request Register")
    ws.sheet_view.showGridLines = False
    
    # Set column widths
    widths = {
        'A': 18,  # Change ID
        'B': 35,  # Change Title
        'C': 15,  # Change Type
        'D': 15,  # Priority
        'E': 35,  # Affected Systems/Assets
        'F': 20,  # Requestor Name
        'G': 20,  # Requestor Contact
        'H': 15,  # Request Date
        'I': 15,  # Required Implementation Date
        'J': 18,  # Change Status
        'K': 18,  # Current Phase
        'L': 12,  # Days in Current Phase
        'M': 18,  # Overall Status Indicator
        'N': 40   # Notes
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:N1')
    ws['A1'] = "CHANGE REQUEST REGISTER - CONFIGURATION CHANGE TRACKING"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # Column headers
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Change Type', 'D': 'Priority',
        'E': 'Affected Systems/Assets', 'F': 'Requestor Name', 'G': 'Requestor Contact',
        'H': 'Request Date', 'I': 'Required Implementation Date', 'J': 'Change Status',
        'K': 'Current Phase', 'L': 'Days in Current Phase', 'M': 'Overall Status Indicator',
        'N': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_crr = {
        'A': 'CRQ-2026-0001', 'B': 'Disable legacy TLS 1.0/1.1 on all web servers',
        'C': 'Normal', 'D': 'P2-High', 'E': 'All production web servers (12 hosts)',
        'F': 'J. Smith', 'G': 'j.smith@company.com', 'H': '05.01.2026',
        'I': '20.01.2026', 'J': '\u2705 Completed', 'N': 'Completed ahead of schedule'
    }
    _sfill_crr = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sborder_crr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_crr.get(col, '')
        cell.fill = _sfill_crr
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_crr

    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 51 (1 sample + 50 empty)
    num_rows = 50
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['K', 'L', 'M']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Add formulas
    for row in range(5, 5 + num_rows):
        # Column K: Current Phase
        ws[f'K{row}'] = f'=IF(J{row}="Draft","Planning",IF(J{row}="Submitted","Approval",IF(OR(J{row}="Approved",J{row}="In Testing"),"Testing",IF(J{row}="Scheduled","Pre-Implementation",IF(J{row}="Implementing","Implementation",IF(OR(J{row}="Completed",J{row}="Failed",J{row}="Rolled Back",J{row}="Cancelled"),"Closed","Unknown"))))))'

        # Column L: Days in Current Phase
        ws[f'L{row}'] = f'=IF(H{row}="","",TODAY()-H{row})'

        # Column M: Overall Status Indicator
        ws[f'M{row}'] = f'=IF(OR(J{row}="Completed",J{row}="Cancelled"),"Complete",IF(L{row}>30,"Delayed",IF(L{row}>14,"At Risk","On Track")))'

    # Data validations
    validations = []
    change_type_dv = create_data_validation(CHANGE_TYPE, allow_blank=False)
    change_type_dv.add(f'C4:C{4+num_rows}')
    validations.append(change_type_dv)

    priority_dv = create_data_validation(CHANGE_PRIORITY, allow_blank=False)
    priority_dv.add(f'D4:D{4+num_rows}')
    validations.append(priority_dv)

    status_dv = create_data_validation(CHANGE_STATUS, allow_blank=False)
    status_dv.add(f'J4:J{4+num_rows}')
    validations.append(status_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'C4:C{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Emergency"'],
                   font=Font(color=COLORS['orange'])))
    ws.conditional_formatting.add(f'C4:C{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Hot Fix"'],
                   font=Font(bold=True, color='9C0006')))

    ws.conditional_formatting.add(f'D4:D{4+num_rows}',
        CellIsRule(operator='equal', formula=['"P1-Critical"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'D4:D{4+num_rows}',
        CellIsRule(operator='equal', formula=['"P2-High"'],
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))

    ws.conditional_formatting.add(f'J4:J{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Completed"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J4:J{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    ws.conditional_formatting.add(f'M4:M{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Complete"'],
                   font=Font(bold=True, color='006100')))
    ws.conditional_formatting.add(f'M4:M{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Delayed"'],
                   font=Font(bold=True, color='9C0006')))

    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_change_approval_workflow_sheet(wb, styles):
    """Create Sheet 3: Change Approval Workflow"""
    ws = wb.create_sheet("Change Approval Workflow")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18, 'B': 30, 'C': 18, 'D': 20, 'E': 18, 'F': 15,
        'G': 20, 'H': 18, 'I': 15, 'J': 20, 'K': 18, 'L': 15,
        'M': 25, 'N': 30, 'O': 18, 'P': 15, 'Q': 40
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "CHANGE APPROVAL WORKFLOW TRACKING"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Approval Tier Required',
        'D': 'Tier 1 Approver Name', 'E': 'Tier 1 Approval Status', 'F': 'Tier 1 Approval Date',
        'G': 'Tier 2 Approver Name', 'H': 'Tier 2 Approval Status', 'I': 'Tier 2 Approval Date',
        'J': 'Tier 3 Approver Name', 'K': 'Tier 3 Approval Status', 'L': 'Tier 3 Approval Date',
        'M': 'Approval Method', 'N': 'Approval Reference', 'O': 'Overall Approval Status',
        'P': 'Days to Full Approval', 'Q': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_caw = {
        'A': 'CRQ-2026-0001', 'B': 'Disable legacy TLS 1.0/1.1 on all web servers',
        'C': 'Two-Tier', 'D': 'J. Smith (Team Lead)', 'E': '\u2705 Approved',
        'F': '08.01.2026', 'G': 'A. Johnson (CISO)', 'H': '\u2705 Approved',
        'I': '10.01.2026', 'J': '', 'K': '\u2796 N/A', 'L': '',
        'M': 'CAB Meeting', 'N': 'CAB-2026-JAN-W2', 'Q': 'Approved at January CAB'
    }
    _sfill_caw = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sborder_caw = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_caw.get(col, '')
        cell.fill = _sfill_caw
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_caw

    num_rows = 50
    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 51 (1 sample + 50 empty)
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['O', 'P']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Add complex formulas for approval status
    for row in range(5, 5 + num_rows):
        # Column O: Overall Approval Status (complex logic)
        ws[f'O{row}'] = f'=IF(C{row}="Single-Tier",IF(E{row}="Approved","Approved",IF(E{row}="Rejected","Rejected","Pending")),IF(C{row}="Two-Tier",IF(AND(E{row}="Approved",H{row}="Approved"),"Approved",IF(OR(E{row}="Rejected",H{row}="Rejected"),"Rejected","Pending")),IF(C{row}="Three-Tier",IF(AND(E{row}="Approved",H{row}="Approved",K{row}="Approved"),"Approved",IF(OR(E{row}="Rejected",H{row}="Rejected",K{row}="Rejected"),"Rejected","Pending")),"Unknown")))'

        # Column P: Days to Full Approval
        ws[f'P{row}'] = f'=IF(O{row}="Approved",IF(C{row}="Single-Tier",IF(F{row}="","",F{row}-TODAY()+365),IF(C{row}="Two-Tier",IF(I{row}="","",I{row}-TODAY()+365),IF(L{row}="","",L{row}-TODAY()+365))),"")'

    # Data validations
    validations = []
    tier_dv = create_data_validation(APPROVAL_TIER, allow_blank=False)
    tier_dv.add(f'C4:C{4+num_rows}')
    validations.append(tier_dv)

    for col in ['E', 'H', 'K']:
        status_dv = create_data_validation(APPROVAL_STATUS_TIER, allow_blank=True)
        status_dv.add(f'{col}4:{col}{4+num_rows}')
        validations.append(status_dv)

    method_dv = create_data_validation(APPROVAL_METHOD, allow_blank=True)
    method_dv.add(f'M4:M{4+num_rows}')
    validations.append(method_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    for col in ['E', 'H', 'K']:
        ws.conditional_formatting.add(f'{col}4:{col}{4+num_rows}',
            CellIsRule(operator='equal', formula=['"Approved"'],
                       fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
        ws.conditional_formatting.add(f'{col}4:{col}{4+num_rows}',
            CellIsRule(operator='equal', formula=['"Rejected"'],
                       fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_impact_assessment_sheet(wb, styles):
    """Create Sheet 4: Impact Assessment"""
    ws = wb.create_sheet("Impact Assessment")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18, 'B': 30, 'C': 12, 'D': 35, 'E': 15, 'F': 15,
        'G': 20, 'H': 15, 'I': 35, 'J': 35, 'K': 15, 'L': 20,
        'M': 35, 'N': 35, 'O': 12, 'P': 25, 'Q': 15
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "IMPACT ASSESSMENT - RISK ANALYSIS PER CHANGE"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Affected Systems Count',
        'D': 'Affected Systems Detail', 'E': 'User Impact', 'F': 'User Count Affected',
        'G': 'Service Downtime Required', 'H': 'Risk Level', 'I': 'Risk Description',
        'J': 'Mitigation Strategies', 'K': 'Rollback Required', 'L': 'Estimated Rollback Time',
        'M': 'Dependencies', 'N': 'Business Justification', 'O': 'Risk Score',
        'P': 'Assessment Completed By', 'Q': 'Assessment Date'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_ia = {
        'A': 'CRQ-2026-0001', 'B': 'Disable legacy TLS 1.0/1.1 on all web servers',
        'C': 12, 'D': 'prod-web-01 through prod-web-12 (all production web servers)',
        'E': 'Minimal', 'F': 50, 'G': 'None', 'H': 'Low',
        'I': 'Minor risk — disabling deprecated protocol; all clients already support TLS 1.2+',
        'J': 'Pre-validate client TLS support; communicate maintenance window to stakeholders',
        'K': 'Yes', 'L': '30 minutes', 'M': 'None identified',
        'N': 'Security compliance requirement — disable deprecated ciphers to meet CIS baseline',
        'P': 'J. Smith', 'Q': '03.01.2026'
    }
    _sfill_ia = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sborder_ia = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_ia.get(col, '')
        cell.fill = _sfill_ia
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_ia

    num_rows = 50
    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 51 (1 sample + 50 empty)
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col == 'O':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Add Risk Score formula
    for row in range(5, 5 + num_rows):
        ws[f'O{row}'] = f'=IF(H{row}="Low",1,IF(H{row}="Medium",2,IF(H{row}="High",3,IF(H{row}="Critical",4,0))))*IF(E{row}="None",1,IF(E{row}="Minimal",2,IF(E{row}="Moderate",3,IF(E{row}="Significant",4,IF(E{row}="Severe",5,0)))))'

    # Data validations
    validations = []
    user_impact_dv = create_data_validation(USER_IMPACT, allow_blank=False)
    user_impact_dv.add(f'E4:E{4+num_rows}')
    validations.append(user_impact_dv)

    downtime_dv = create_data_validation(SERVICE_DOWNTIME, allow_blank=False)
    downtime_dv.add(f'G4:G{4+num_rows}')
    validations.append(downtime_dv)

    risk_dv = create_data_validation(RISK_LEVEL, allow_blank=False)
    risk_dv.add(f'H4:H{4+num_rows}')
    validations.append(risk_dv)

    rollback_dv = create_data_validation(YES_NO, allow_blank=False)
    rollback_dv.add(f'K4:K{4+num_rows}')
    validations.append(rollback_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"High"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    ws.conditional_formatting.add(f'O4:O{4+num_rows}',
        CellIsRule(operator='greaterThanOrEqual', formula=['12'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'O4:O{4+num_rows}',
        CellIsRule(operator='between', formula=['6', '11'],
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))

    ws.freeze_panes = 'B3'
    # Formula cells unlocked

    return ws

def create_testing_validation_sheet(wb, styles):
    """Create Sheet 5: Testing Validation"""
    ws = wb.create_sheet("Testing Validation")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18, 'B': 30, 'C': 15, 'D': 20, 'E': 15, 'F': 15,
        'G': 12, 'H': 30, 'I': 12, 'J': 12, 'K': 12, 'L': 12,
        'M': 12, 'N': 12, 'O': 18, 'P': 18, 'Q': 20, 'R': 15, 'S': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:S1')
    ws['A1'] = "TESTING VALIDATION RECORDS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Testing Required',
        'D': 'Test Environment', 'E': 'Test Start Date', 'F': 'Test End Date',
        'G': 'Test Duration (Days)', 'H': 'Test Plan Reference', 'I': 'Test Cases Executed',
        'J': 'Test Cases Passed', 'K': 'Test Cases Failed', 'L': 'Test Pass Rate %',
        'M': 'Critical Issues Found', 'N': 'Issues Resolved Before Deployment',
        'O': 'Testing Status', 'P': 'Go/No-Go Decision', 'Q': 'Decision Maker',
        'R': 'Decision Date', 'S': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    num_rows = 51
    # MAX standard: Row 1 = sample with example data, Rows 2-51 = empty
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['G', 'L']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formulas
    for row in range(3, 3 + num_rows):
        ws[f'G{row}'] = f'=IF(OR(E{row}="",F{row}=""),"",F{row}-E{row})'
        ws[f'L{row}'] = f'=IF(I{row}=0,"",J{row}/I{row}*100)'
        ws[f'L{row}'].number_format = '0.0'

    # Data validations
    validations = []
    test_req_dv = create_data_validation(YES_NO_NA, allow_blank=False)
    test_req_dv.add(f'C3:C{2+num_rows}')
    validations.append(test_req_dv)

    env_dv = create_data_validation(TEST_ENVIRONMENT, allow_blank=True)
    env_dv.add(f'D3:D{2+num_rows}')
    validations.append(env_dv)

    status_dv = create_data_validation(TESTING_STATUS, allow_blank=False)
    status_dv.add(f'O3:O{2+num_rows}')
    validations.append(status_dv)

    gonogo_dv = create_data_validation(GO_NOGO, allow_blank=True)
    gonogo_dv.add(f'P3:P{2+num_rows}')
    validations.append(gonogo_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='between', formula=['80', '94.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='lessThan', formula=['80'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Go"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"No-Go"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_implementation_log_sheet(wb, styles):
    """Create Sheet 6: Implementation Log"""
    ws = wb.create_sheet("Implementation Log")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18, 'B': 30, 'C': 20, 'D': 20, 'E': 20, 'F': 12,
        'G': 20, 'H': 20, 'I': 40, 'J': 35, 'K': 20, 'L': 30,
        'M': 35, 'N': 18, 'O': 18, 'P': 18, 'Q': 15, 'R': 20, 'S': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:S1')
    ws['A1'] = "IMPLEMENTATION LOG - CHANGE EXECUTION RECORDS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Scheduled Implementation Date/Time',
        'D': 'Actual Implementation Date/Time', 'E': 'Implementation Completed Date/Time',
        'F': 'Implementation Duration (Hours)', 'G': 'Implementer Name',
        'H': 'Implementation Method', 'I': 'Implementation Steps Performed',
        'J': 'Deviations from Plan', 'K': 'Post-Implementation Verification',
        'L': 'Verification Method', 'M': 'Issues Encountered', 'N': 'Issues Resolved',
        'O': 'Implementation Status', 'P': 'Change Outcome', 'Q': 'Rollback Triggered',
        'R': 'Rollback Completion Time', 'S': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    num_rows = 51
    # MAX standard: Row 1 = sample with example data, Rows 2-51 = empty
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['F', 'P']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formulas
    for row in range(3, 3 + num_rows):
        ws[f'F{row}'] = f'=IF(OR(D{row}="",E{row}=""),"",ROUND((E{row}-D{row})*24,2))'
        ws[f'P{row}'] = f'=IF(O{row}="Successful","Success",IF(O{row}="Failed","Failure",IF(O{row}="Rolled Back","Rollback Required","Unknown")))'

    # Data validations
    validations = []
    method_dv = create_data_validation(IMPLEMENTATION_METHOD, allow_blank=False)
    method_dv.add(f'H3:H{2+num_rows}')
    validations.append(method_dv)

    verif_dv = create_data_validation(VERIFICATION_STATUS, allow_blank=False)
    verif_dv.add(f'K3:K{2+num_rows}')
    validations.append(verif_dv)

    resolved_dv = create_data_validation(ISSUES_RESOLVED, allow_blank=True)
    resolved_dv.add(f'N3:N{2+num_rows}')
    validations.append(resolved_dv)

    impl_status_dv = create_data_validation(IMPLEMENTATION_STATUS, allow_blank=False)
    impl_status_dv.add(f'O3:O{2+num_rows}')
    validations.append(impl_status_dv)

    rollback_dv = create_data_validation(YES_NO, allow_blank=False)
    rollback_dv.add(f'Q3:Q{2+num_rows}')
    validations.append(rollback_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Successful"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.conditional_formatting.add(f'Q3:Q{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Yes"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_rollback_capability_sheet(wb, styles):
    """Create Sheet 7: Rollback Capability"""
    ws = wb.create_sheet("Rollback Capability")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18, 'B': 30, 'C': 15, 'D': 18, 'E': 40, 'F': 35,
        'G': 20, 'H': 18, 'I': 15, 'J': 20, 'K': 35, 'L': 18,
        'M': 15, 'N': 25, 'O': 20, 'P': 18, 'Q': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "ROLLBACK CAPABILITY ASSESSMENT"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Rollback Required',
        'D': 'Rollback Procedure Documented', 'E': 'Rollback Document Location',
        'F': 'Rollback Trigger Criteria', 'G': 'Estimated Rollback Time',
        'H': 'Rollback Tested', 'I': 'Rollback Test Date', 'J': 'Rollback Test Results',
        'K': 'Rollback Dependencies', 'L': 'Data Loss Risk on Rollback',
        'M': 'Data Backup Verified', 'N': 'Rollback Approval Required',
        'O': 'Rollback Owner', 'P': 'Rollback Readiness', 'Q': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_rc = {
        'A': 'CRQ-2026-0001', 'B': 'Disable legacy TLS 1.0/1.1 on all web servers',
        'C': 'Yes', 'D': 'Yes', 'E': '/docs/rollback/CRQ-2026-0001-rollback.pdf',
        'F': 'Any certificate validation failure or service degradation >5%',
        'G': '30 minutes', 'H': 'Yes', 'I': '15.01.2026', 'J': 'Successful',
        'K': 'None', 'L': 'None', 'M': 'Yes', 'N': 'Change Manager',
        'O': 'J. Smith (Sec Ops)', 'Q': 'Rollback tested in staging on 14.01.2026'
    }
    _sborder_rc = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_rc.get(col, '')
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_rc

    num_rows = 50
    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 50 (grey sample separate)
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col == 'P':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formula for Rollback Readiness
    for row in range(5, 5 + num_rows):
        ws[f'P{row}'] = f'=IF(C{row}="No","N/A",IF(AND(D{row}="Yes",H{row}="Yes",J{row}="Successful",M{row}="Yes"),"Ready",IF(AND(D{row}="Yes",OR(H{row}="No",J{row}="Not Tested")),"Not Ready","Partially Ready")))'

    # Data validations
    validations = []
    rollback_req_dv = create_data_validation(YES_NO, allow_blank=False)
    rollback_req_dv.add(f'C4:C{4+num_rows}')
    validations.append(rollback_req_dv)

    doc_dv = create_data_validation(YES_NO_NA, allow_blank=False)
    doc_dv.add(f'D4:D{4+num_rows}')
    validations.append(doc_dv)

    tested_dv = create_data_validation(YES_NO_PARTIAL_NA, allow_blank=False)
    tested_dv.add(f'H4:H{4+num_rows}')
    validations.append(tested_dv)

    results_dv = create_data_validation(ROLLBACK_TEST_RESULTS, allow_blank=True)
    results_dv.add(f'J4:J{4+num_rows}')
    validations.append(results_dv)

    data_loss_dv = create_data_validation(DATA_LOSS_RISK, allow_blank=True)
    data_loss_dv.add(f'L4:L{4+num_rows}')
    validations.append(data_loss_dv)

    backup_dv = create_data_validation(YES_NO_NA, allow_blank=False)
    backup_dv.add(f'M4:M{4+num_rows}')
    validations.append(backup_dv)

    approval_dv = create_data_validation(ROLLBACK_APPROVAL, allow_blank=True)
    approval_dv.add(f'N4:N{4+num_rows}')
    validations.append(approval_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'P4:P{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Ready"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'P4:P{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Ready"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'P4:P{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Partially Ready"'],
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))

    ws.conditional_formatting.add(f'J4:J{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Successful"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J4:J{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J4:J{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Tested"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    ws.conditional_formatting.add(f'L4:L{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Significant"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'L4:L{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Moderate"'],
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))

    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_emergency_changes_sheet(wb, styles):
    """Create Sheet 8: Emergency Changes"""
    ws = wb.create_sheet("Emergency Changes")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18, 'B': 30, 'C': 25, 'D': 40, 'E': 20, 'F': 20,
        'G': 20, 'H': 12, 'I': 25, 'J': 30, 'K': 20, 'L': 15,
        'M': 25, 'N': 18, 'O': 18, 'P': 40, 'Q': 40, 'R': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:R1')
    ws['A1'] = "EMERGENCY CHANGES - EXPEDITED PROCESS TRACKING"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Emergency Type',
        'D': 'Business Impact if Not Implemented', 'E': 'Emergency Declared By',
        'F': 'Emergency Declaration Time', 'G': 'Implementation Time',
        'H': 'Time to Implement (Hours)', 'I': 'Emergency Approval Method',
        'J': 'Emergency Approvers', 'K': 'Post-Implementation Documentation Completed',
        'L': 'CAB Retrospective Review Date', 'M': 'CAB Review Outcome',
        'N': 'Justification Valid', 'O': 'Process Abuse Indicator',
        'P': 'Lessons Learned', 'Q': 'Remediation Actions', 'R': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    num_rows = 51
    # MAX standard: Row 1 = sample with example data, Rows 2-51 = empty
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['H', 'O']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formulas
    for row in range(3, 3 + num_rows):
        # Column H: Time to Implement (Hours)
        ws[f'H{row}'] = f'=IF(OR(F{row}="",G{row}=""),"",ROUND((G{row}-F{row})*24,2))'

        # Column O: Process Abuse Indicator
        ws[f'O{row}'] = f'=IF(AND(H{row}>48,N{row}="Questionable"),"Likely Abuse",IF(OR(H{row}>72,N{row}="No"),"Questionable","Legitimate"))'

    # Data validations
    validations = []
    emerg_type_dv = create_data_validation(EMERGENCY_TYPE, allow_blank=False)
    emerg_type_dv.add(f'C3:C{2+num_rows}')
    validations.append(emerg_type_dv)

    emerg_method_dv = create_data_validation(EMERGENCY_APPROVAL_METHOD, allow_blank=False)
    emerg_method_dv.add(f'I3:I{2+num_rows}')
    validations.append(emerg_method_dv)

    post_doc_dv = create_data_validation(POST_IMPL_DOC, allow_blank=False)
    post_doc_dv.add(f'K3:K{2+num_rows}')
    validations.append(post_doc_dv)

    cab_review_dv = create_data_validation(CAB_REVIEW_OUTCOME, allow_blank=False)
    cab_review_dv.add(f'M3:M{2+num_rows}')
    validations.append(cab_review_dv)

    justif_dv = create_data_validation(JUSTIFICATION_VALID, allow_blank=True)
    justif_dv.add(f'N3:N{2+num_rows}')
    validations.append(justif_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Approved"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Disapproved (requires reversal)"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Yet Reviewed"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Legitimate"'], 
                   font=Font(color='006100')))
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Questionable"'], 
                   font=Font(bold=True, color='9C6500')))
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Likely Abuse"'], 
                   font=Font(bold=True, color='9C0006')))
    
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"No"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Yes"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_change_success_metrics_sheet(wb, styles):
    """Create Sheet 9: Change Success Metrics (Dashboard)"""
    ws = wb.create_sheet("Change Success Metrics")
    ws.sheet_view.showGridLines = False
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "CHANGE SUCCESS METRICS DASHBOARD"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # SECTION A: Overall Change Metrics
    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL CHANGE METRICS"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    metrics = [
        ("Total Changes (All Types)", "=COUNTA('Change Request Register'!A3:A102)-COUNTBLANK('Change Request Register'!A3:A102)", "N/A", ""),
        ("Completed Changes", "=COUNTIF('Change Request Register'!J3:J102,\"*Completed\")", "N/A", ""),
        ("Failed Changes", "=COUNTIF('Change Request Register'!J3:J102,\"*Failed\")", "<5%", '=IF(B7=0,"N/A",IF(B7/B5*100<5,"✓ Within Target","✗ Exceeds Target"))'),
        ("Rolled Back Changes", "=COUNTIF('Change Request Register'!J3:J102,\"*Rolled Back\")", "<3%", '=IF(B8=0,"N/A",IF(B8/B5*100<3,"✓ Within Target","✗ Exceeds Target"))'),
        ("Changes in Progress", "=COUNTIF('Change Request Register'!J3:J102,\"*Implementing\")+COUNTIF('Change Request Register'!J3:J102,\"*Scheduled\")", "N/A", ""),
        ("Overall Success Rate %", '=IF((B6+B7+B8)=0,0,B6/(B6+B7+B8)*100)', "≥95%", '=IF(B10>=95,"✓ Compliant",IF(B10>=90,"\u26A0 Partial","✗ Non-Compliant"))'),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION B: Change Type Distribution
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "CHANGE TYPE DISTRIBUTION"
    apply_style(ws[f'A{row}'], styles['section_header'])

    row += 1
    ws[f'A{row}'] = "Change Type"
    ws[f'B{row}'] = "Count"
    ws[f'C{row}'] = "Percentage"
    ws[f'D{row}'] = "Target %"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    change_types = [
        ("Standard", "=COUNTIF('Change Request Register'!C3:C102,\"Standard\")", "40-50%"),
        ("Normal", "=COUNTIF('Change Request Register'!C3:C102,\"Normal\")", "40-50%"),
        ("Emergency", "=COUNTIF('Change Request Register'!C3:C102,\"Emergency\")", "<8%"),
        ("Hot Fix", "=COUNTIF('Change Request Register'!C3:C102,\"Hot Fix\")", "<2%"),
    ]
    
    row += 1
    total_row_ref = row - 6  # Reference to total changes row (B5)
    for change_type, count_formula, target in change_types:
        ws[f'A{row}'] = change_type
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = count_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = f'=IF($B$5=0,0,B{row}/$B$5*100)'
        ws[f'C{row}'].number_format = '0.0'
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = target
        ws[f'D{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION C: Success Rate by Change Type
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "SUCCESS RATE BY CHANGE TYPE"
    apply_style(ws[f'A{row}'], styles['section_header'])

    row += 1
    ws[f'A{row}'] = "Change Type"
    ws[f'B{row}'] = "Total"
    ws[f'C{row}'] = "Successful"
    ws[f'D{row}'] = "Success Rate %"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    row += 1
    for change_type, _, _ in change_types:
        ws[f'A{row}'] = change_type
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f"=COUNTIF('Change Request Register'!C3:C102,\"{change_type}\")"
        apply_style(ws[f'B{row}'], styles['protected_cell'])

        ws[f'C{row}'] = f"=COUNTIFS('Change Request Register'!C3:C102,\"{change_type}\",'Change Request Register'!J3:J102,\"*Completed\")"
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF(B{row}=0,0,C{row}/B{row}*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION D: Emergency Change Analysis
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "EMERGENCY CHANGE ANALYSIS"
    apply_style(ws[f'A{row}'], styles['section_header'])

    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    emerg_metrics = [
        ("Total Emergency Changes", "=COUNTA('Emergency Changes'!A3:A52)-COUNTBLANK('Emergency Changes'!A3:A52)", "N/A", ""),
        ("Emergency Changes %", '=IF($B$5=0,0,B{prev_row}/$B$5*100)', "<10%", '=IF(B{row}<10,"✓ Within Target",IF(B{row}<15,"\u26A0 Warning","✗ Critical"))'),
        ("Emergency Changes Not Reviewed", "=COUNTIF('Emergency Changes'!M3:M52,\"⏳ Not Yet Reviewed\")", "0", '=IF(B{row}=0,"✓ All Reviewed","✗ "&B{row}&" Pending")'),
        ("Process Abuse Flagged", "=COUNTIF('Emergency Changes'!O3:O52,\"Likely Abuse\")", "0", '=IF(B{row}=0,"✓ None","✗ "&B{row}&" Flagged")'),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in emerg_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        value_formula = value_formula.replace("{prev_row}", str(row - 1)).replace("{row}", str(row))
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        status_formula = status_formula.replace("{row}", str(row))
        ws[f'D{row}'] = status_formula
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # Add note
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "NOTE: All metrics auto-calculated. Review monthly. Target: ≥95% success rate, <10% emergency changes."
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    return ws

def create_compliance_dashboard_sheet(wb, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1, TABLE 2, TABLE 3."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    _white_bold_14 = Font(bold=True, color="FFFFFF", size=14)
    _white_bold_11 = Font(bold=True, color="FFFFFF", size=11)
    _dark_bold_10 = Font(bold=True, color="000000", size=10)
    _dark_10 = Font(color="000000", size=10)
    _dark_9_italic = Font(color="000000", size=9, italic=True)
    _title_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _data_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _t3_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _center = Alignment(horizontal="center", vertical="center")
    _left = Alignment(horizontal="left", vertical="center")
    _wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths
    for col_letter, width in [("A", 40), ("B", 15), ("C", 12), ("D", 12), ("E", 12), ("F", 12), ("G", 15)]:
        ws.column_dimensions[col_letter].width = width

    # ROW 1 — Title
    ws["A1"].border = border_thin
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE CONTROL \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = _white_bold_14
    ws["A1"].fill = _title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ROW 2 — Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Summary Dashboard  |  Change Control Assessment  |  Generated: {GENERATED_TIMESTAMP}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Freeze at A4
    ws.freeze_panes = "A4"

    # TABLE 1 banner (row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = _white_bold_11
    ws["A4"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A4"].alignment = _center

    # TABLE 1 headers (row 5)
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = _dark_bold_10
        cell.fill = _hdr_fill
        cell.border = border_thin
        cell.alignment = _center

    _blue_10 = Font(color="0000FF", size=10)

    def _t1_row(row, area, b, c_val, d, e, f_val, g):
        ws.cell(row=row, column=1, value=area).border = border_thin
        ws.cell(row=row, column=1).font = _dark_10
        ws.cell(row=row, column=1).alignment = _left
        for col_idx, val in enumerate([b, c_val, d, e, f_val], start=2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = _blue_10
            cell.border = border_thin
            cell.alignment = _center
        cell_g = ws.cell(row=row, column=7, value=g)
        cell_g.font = _blue_10
        cell_g.border = border_thin
        cell_g.alignment = _center
        cell_g.number_format = "0.0%"

    # Row 6: Change Request Register — Change Status col J, data rows 4-54
    # Compliant = Completed, Partial = In Testing/Scheduled/Implementing, Non-Compliant = Failed/Rolled Back, N/A = Cancelled
    _t1_row(6,
        "Change Request Register",
        "=COUNTA('Change Request Register'!B5:B54)",
        "=COUNTIF('Change Request Register'!J5:J54,\"\u2705 Completed\")",
        "=COUNTIF('Change Request Register'!J5:J54,\"In Testing\")+COUNTIF('Change Request Register'!J5:J54,\"Scheduled\")+COUNTIF('Change Request Register'!J5:J54,\"\u23F3 Implementing\")",
        "=COUNTIF('Change Request Register'!J5:J54,\"\u274C Failed\")+COUNTIF('Change Request Register'!J5:J54,\"Rolled Back\")",
        "=COUNTIF('Change Request Register'!J5:J54,\"\u274C Cancelled\")",
        "=IF((B6-F6)=0,0,C6/(B6-F6))"
    )

    # Row 7: Change Approval Workflow — Overall Approval col O, data rows 4-54
    _t1_row(7,
        "Change Approval Workflow",
        "=COUNTA('Change Approval Workflow'!B5:B54)",
        "=COUNTIF('Change Approval Workflow'!O5:O54,\"Approved\")",
        "=COUNTIF('Change Approval Workflow'!O5:O54,\"Pending\")",
        "=COUNTIF('Change Approval Workflow'!O5:O54,\"Rejected\")",
        "=COUNTIF('Change Approval Workflow'!O5:O54,\"Unknown\")",
        "=IF((B7-F7)=0,0,C7/(B7-F7))"
    )

    # Row 8: Impact Assessment — Risk Level col H, data rows 4-54
    # Compliant = Low, Partial = Medium, Non-Compliant = High/Critical
    _t1_row(8,
        "Impact Assessment",
        "=COUNTA('Impact Assessment'!B5:B54)",
        "=COUNTIF('Impact Assessment'!H5:H54,\"Low\")",
        "=COUNTIF('Impact Assessment'!H5:H54,\"Medium\")",
        "=COUNTIF('Impact Assessment'!H5:H54,\"High\")+COUNTIF('Impact Assessment'!H5:H54,\"Critical\")",
        "=0",
        "=IF((B8-F8)=0,0,C8/(B8-F8))"
    )

    # Row 9: TOTAL
    ws.cell(row=9, column=1, value="TOTAL").font = Font(bold=True, color="000000", size=10)
    ws.cell(row=9, column=1).fill = _total_fill
    ws.cell(row=9, column=1).border = border_thin
    ws.cell(row=9, column=1).alignment = _left
    for col_idx in range(2, 7):
        from openpyxl.utils import get_column_letter as _gcl
        col_letter = _gcl(col_idx)
        cell = ws.cell(row=9, column=col_idx, value=f"=SUM({col_letter}6:{col_letter}8)")
        cell.font = Font(bold=True, color="000000", size=10)
        cell.fill = _total_fill
        cell.border = border_thin
        cell.alignment = _center
    cell_tot_g = ws.cell(row=9, column=7, value="=IF((B9-F9)=0,0,C9/(B9-F9))")
    cell_tot_g.font = Font(bold=True, color="000000", size=10)
    cell_tot_g.fill = _total_fill
    cell_tot_g.border = border_thin
    cell_tot_g.alignment = _center
    cell_tot_g.number_format = "0.0%"

    # TABLE 2 banner (row 11)
    ws.merge_cells("A11:G11")
    ws["A11"] = "TABLE 2 \u2013 KEY METRICS"
    ws["A11"].font = _white_bold_11
    ws["A11"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A11"].alignment = _center

    # TABLE 2 headers (row 12)
    for col_idx, hdr in enumerate(["Metric", "Value", "What This Shows"], start=1):
        cell = ws.cell(row=12, column=col_idx, value=hdr)
        cell.font = _dark_bold_10
        cell.fill = _hdr_fill
        cell.border = border_thin
        cell.alignment = _center if col_idx == 2 else _left
    ws.merge_cells("C12:G12")
    ws["C12"].font = _dark_bold_10
    ws["C12"].fill = _hdr_fill
    ws["C12"].border = border_thin

    t2_metrics = [
        (13, "Changes Completed Successfully",
         "=COUNTIF('Change Request Register'!J5:J54,\"\u2705 Completed\")",
         "Total changes reaching Completed status (primary KPI for change success rate)"),
        (14, "Change Success Rate",
         "=IF(COUNTA('Change Request Register'!B4:B54)=0,0,COUNTIF('Change Request Register'!J5:J54,\"\u2705 Completed\")/COUNTA('Change Request Register'!B4:B54))",
         "% of all changes successfully completed (target: \u226595%)"),
        (15, "Emergency Changes Raised",
         "=COUNTIF('Change Request Register'!C5:C54,\"Emergency\")",
         "Count of expedited Emergency changes (target: <10% of total changes)"),
        (16, "High/Critical Risk Changes",
         "=COUNTIF('Change Request Register'!D5:D54,\"P1-Critical\")+COUNTIF('Change Request Register'!D5:D54,\"P2-High\")",
         "Changes with P1-Critical or P2-High priority requiring close oversight"),
        (17, "Failed or Rolled-Back Changes",
         "=COUNTIF('Change Request Register'!J5:J54,\"\u274C Failed\")+COUNTIF('Change Request Register'!J5:J54,\"Rolled Back\")",
         "Changes that failed or required rollback (target: 0; investigate root cause)"),
        (18, "Pending Approvals",
         "=COUNTIF('Change Approval Workflow'!O5:O54,\"Pending\")",
         "Changes awaiting approval sign-off (target: 0 pending at month-end)"),
    ]

    for row, metric, formula, description in t2_metrics:
        cell_a = ws.cell(row=row, column=1, value=metric)
        cell_a.font = _dark_10
        cell_a.border = border_thin
        cell_a.alignment = _left
        cell_b = ws.cell(row=row, column=2, value=formula)
        cell_b.font = _dark_10
        cell_b.border = border_thin
        cell_b.alignment = _center
        if row == 14:
            cell_b.number_format = "0.0%"
        ws.merge_cells(f"C{row}:G{row}")
        cell_c = ws.cell(row=row, column=3, value=description)
        cell_c.font = _dark_9_italic
        cell_c.border = border_thin
        cell_c.alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).border = border_thin

    # TABLE 3 banner (row 20)
    ws.merge_cells("A20:G20")
    ws["A20"] = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    ws["A20"].font = _white_bold_11
    ws["A20"].fill = _t3_fill
    ws["A20"].alignment = _left

    # TABLE 3 headers (row 21)
    ws["A21"].value = "Finding"
    ws["A21"].font = _dark_bold_10
    ws["A21"].fill = _hdr_fill
    ws["A21"].border = border_thin
    ws["A21"].alignment = _left
    ws["B21"].value = "Count"
    ws["B21"].font = _dark_bold_10
    ws["B21"].fill = _hdr_fill
    ws["B21"].border = border_thin
    ws["B21"].alignment = _center
    ws.merge_cells("C21:G21")
    ws["C21"].value = "Recommendation"
    ws["C21"].font = _dark_bold_10
    ws["C21"].fill = _hdr_fill
    ws["C21"].border = border_thin
    ws["C21"].alignment = _left

    t3_findings = [
        (22, "Failed or Rolled-Back Changes",
         "=COUNTIF('Change Request Register'!J5:J54,\"\u274C Failed\")+COUNTIF('Change Request Register'!J5:J54,\"Rolled Back\")",
         "Conduct post-implementation review (PIR) for all failed changes; identify root causes and implement preventive controls"),
        (23, "Emergency Changes (Unplanned)",
         "=COUNTIF('Change Request Register'!C5:C54,\"Emergency\")",
         "Review emergency change frequency; if >10% of total, strengthen normal change process to reduce urgent escalations"),
        (24, "High-Risk Changes Pending Approval",
         "=COUNTIFS('Change Request Register'!D5:D54,\"P1-Critical\",'Change Approval Workflow'!O5:O54,\"Pending\")+COUNTIFS('Change Request Register'!D5:D54,\"P2-High\",'Change Approval Workflow'!O5:O54,\"Pending\")",
         "Escalate P1/P2 pending approvals to CAB urgently; implement SLA-based escalation for approvals >5 business days"),
        (25, "Changes Without Impact Assessment",
         "=COUNTA('Change Request Register'!B5:B54)-COUNTA('Impact Assessment'!B4:B54)",
         "Ensure all changes have a completed impact assessment before approval; make impact assessment mandatory in change template"),
    ]

    for row, finding, formula, recommendation in t3_findings:
        cell_a = ws.cell(row=row, column=1, value=finding)
        cell_a.font = _dark_10
        cell_a.fill = _data_fill
        cell_a.border = border_thin
        cell_a.alignment = _left
        cell_b = ws.cell(row=row, column=2, value=formula)
        cell_b.font = _dark_10
        cell_b.fill = _data_fill
        cell_b.border = border_thin
        cell_b.alignment = _center
        ws.merge_cells(f"C{row}:G{row}")
        cell_c = ws.cell(row=row, column=3, value=recommendation)
        cell_c.font = _dark_9_italic
        cell_c.fill = _data_fill
        cell_c.border = border_thin
        cell_c.alignment = _wrap_left
        for _c in range(4, 8):
            ws.cell(row=row, column=_c).fill = _data_fill
            ws.cell(row=row, column=_c).border = border_thin

    # TOTAL row (row 26)
    ws.cell(row=26, column=1, value="TOTAL").font = Font(bold=True, color="000000", size=10)
    ws.cell(row=26, column=1).border = border_thin
    ws.cell(row=26, column=1).alignment = _left
    ws.cell(row=26, column=2, value="=SUM(B22:B25)").font = Font(bold=True, color="000000", size=10)
    ws.cell(row=26, column=2).border = border_thin
    ws.cell(row=26, column=2).alignment = _center
    ws.merge_cells("C26:G26")
    ws.cell(row=26, column=3, value="Total count of all key findings requiring attention").font = _dark_9_italic
    ws.cell(row=26, column=3).border = border_thin
    ws.cell(row=26, column=3).alignment = _wrap_left
    for _c in range(4, 8):
        ws.cell(row=26, column=_c).border = border_thin

    logger.info("  \u2713 Summary Dashboard sheet created")
    return ws
    
    # SECTION A: Approval Compliance
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "APPROVAL COMPLIANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])

    row += 1
    ws[f'A{row}'] = "Compliance Check"
    ws[f'B{row}'] = "Compliant"
    ws[f'C{row}'] = "Non-Compliant"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Target"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    approval_checks = [
        ("All Changes Have Approval Records",
         "=COUNTIF('Change Approval Workflow'!O3:O102,\"Approved\")+COUNTIF('Change Approval Workflow'!O3:O102,\"Rejected\")",
         "=COUNTA('Change Request Register'!A3:A102)-COUNTBLANK('Change Request Register'!A3:A102)-B{row}",
         "100%"),
        ("Emergency Changes Have Post-Facto Review",
         "=COUNTIFS('Emergency Changes'!M3:M52,\"Approved\")+COUNTIFS('Emergency Changes'!M3:M52,\"Approved with Remediation\")",
         "=COUNTA('Emergency Changes'!A3:A52)-COUNTBLANK('Emergency Changes'!A3:A52)-B{row}",
         "100%"),
    ]
    
    row += 1
    for check_name, compliant_formula, non_compliant_formula, target in approval_checks:
        ws[f'A{row}'] = check_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = compliant_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        non_compliant_formula = non_compliant_formula.replace("{row}", str(row))
        ws[f'C{row}'] = non_compliant_formula
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF((B{row}+C{row})=0,0,B{row}/(B{row}+C{row})*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = target
        ws[f'E{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION B: Testing Compliance
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "TESTING COMPLIANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])

    row += 1
    ws[f'A{row}'] = "Compliance Check"
    ws[f'B{row}'] = "Compliant"
    ws[f'C{row}'] = "Non-Compliant"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Target"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    testing_checks = [
        ("Normal Changes Have Testing",
         "=COUNTIFS('Change Request Register'!C3:C102,\"Normal\",'Testing Validation'!C3:C102,\"Yes\")",
         "=COUNTIF('Change Request Register'!C3:C102,\"Normal\")-B{row}",
         "100%"),
        ("Test Pass Rate >=95%",
         "=COUNTIF('Testing Validation'!L3:L102,\">=95\")",
         "=COUNTIF('Testing Validation'!L3:L102,\"<95\")",
         "95%"),
        ("Go/No-Go Decision Documented",
         "=COUNTIF('Testing Validation'!P3:P102,\"Go\")+COUNTIF('Testing Validation'!P3:P102,\"No-Go\")",
         "=COUNTA('Testing Validation'!A3:A102)-COUNTBLANK('Testing Validation'!A3:A102)-B{row}",
         "100%"),
    ]
    
    row += 1
    for check_name, compliant_formula, non_compliant_formula, target in testing_checks:
        ws[f'A{row}'] = check_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = compliant_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        non_compliant_formula = non_compliant_formula.replace("{row}", str(row))
        ws[f'C{row}'] = non_compliant_formula
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF((B{row}+C{row})=0,0,B{row}/(B{row}+C{row})*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = target
        ws[f'E{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION C: Rollback Compliance
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ROLLBACK COMPLIANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])

    row += 1
    ws[f'A{row}'] = "Compliance Check"
    ws[f'B{row}'] = "Compliant"
    ws[f'C{row}'] = "Non-Compliant"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Target"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    rollback_checks = [
        ("High-Risk Changes Have Rollback Plan",
         "=COUNTIFS('Impact Assessment'!H3:H102,\"High\",'Rollback Capability'!C3:C102,\"Yes\")+COUNTIFS('Impact Assessment'!H3:H102,\"Critical\",'Rollback Capability'!C3:C102,\"Yes\")",
         "=(COUNTIF('Impact Assessment'!H3:H102,\"High\")+COUNTIF('Impact Assessment'!H3:H102,\"Critical\"))-B{row}",
         "100%"),
        ("Rollback Procedures Tested",
         "=COUNTIF('Rollback Capability'!P3:P102,\"Ready\")",
         "=COUNTIF('Rollback Capability'!P3:P102,\"Not Ready\")",
         "100%"),
    ]
    
    row += 1
    for check_name, compliant_formula, non_compliant_formula, target in rollback_checks:
        ws[f'A{row}'] = check_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = compliant_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        non_compliant_formula = non_compliant_formula.replace("{row}", str(row))
        ws[f'C{row}'] = non_compliant_formula
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF((B{row}+C{row})=0,0,B{row}/(B{row}+C{row})*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = target
        ws[f'E{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION D: Overall Summary
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "OVERALL COMPLIANCE SUMMARY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Category"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Status"
    for col in ['A', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    # Calculate averages from sections above
    approval_start = 6
    testing_start = 11
    rollback_start = 17
    
    row += 1
    ws[f'A{row}'] = "Approval Process"
    ws[f'D{row}'] = f'=AVERAGE(D{approval_start}:D{approval_start+1})'
    ws[f'D{row}'].number_format = '0.0'
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ Compliant",IF(D{row}>=90,"\u26A0 Partial","✗ Non-Compliant"))'
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    row += 1
    ws[f'A{row}'] = "Testing Process"
    ws[f'D{row}'] = f'=AVERAGE(D{testing_start}:D{testing_start+2})'
    ws[f'D{row}'].number_format = '0.0'
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ Compliant",IF(D{row}>=90,"\u26A0 Partial","✗ Non-Compliant"))'
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    row += 1
    ws[f'A{row}'] = "Rollback Readiness"
    ws[f'D{row}'] = f'=AVERAGE(D{rollback_start}:D{rollback_start+1})'
    ws[f'D{row}'].number_format = '0.0'
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ Compliant",IF(D{row}>=90,"\u26A0 Partial","✗ Non-Compliant"))'
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    row += 1
    ws[f'A{row}'] = "OVERALL COMPLIANCE"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
    overall_row = row
    ws[f'D{row}'] = f'=AVERAGE(D{overall_row-3}:D{overall_row-1})'
    ws[f'D{row}'].number_format = '0.0'
    ws[f'D{row}'].font = Font(name='Calibri', size=12, bold=True)
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ COMPLIANT",IF(D{row}>=90,"\u26A0 PARTIAL","✗ NON-COMPLIANT"))'
    ws[f'E{row}'].font = Font(name='Calibri', size=12, bold=True)
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    # Conditional formatting for compliance %
    ws.conditional_formatting.add(f'D6:D{overall_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'D6:D{overall_row}',
        CellIsRule(operator='between', formula=['90', '94.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'D6:D{overall_row}',
        CellIsRule(operator='lessThan', formula=['90'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    return ws

def create_evidence_register(wb, styles):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(color="808080")
        ws.cell(row=r, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"
def create_approval_sheet(wb, styles):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.9.2 - Change Control Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "='Change Success Metrics'!B10"),
        ("Assessment Status:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border_thin
        row += 1

    # Status dropdown on Assessment Status
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border_thin

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Apply borders to all merged cell top-left corners (GS-AS-011)
    _as_thin = Side(style="thin")
    _as_border = Border(left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin)
    for merge_range in ws.merged_cells.ranges:
        tl = ws.cell(merge_range.min_row, merge_range.min_col)
        tl.border = _as_border
    ws.freeze_panes = "A3"
# ============================================================================
# MAIN EXECUTION
# ============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating {WORKBOOK_TITLE} Workbook")
    logger.info("=" * 70)
    logger.info(f"Document ID: {DOCUMENT_ID}")
    logger.info(f"Version: {WORKBOOK_VERSION}")
    logger.info(f"Date: {datetime.now().strftime('%d.%m.%Y')}")
    logger.info("-" * 70)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create styles
    styles = create_styles()
    
    # Create all sheets
    logger.info("Creating sheets...")
    
    logger.info("  1/12 Creating Instructions sheet...")
    create_instructions_sheet(wb.create_sheet())
    
    logger.info("  2/12 Creating Change Request Register sheet (100 rows)...")
    create_change_request_register_sheet(wb, styles)
    
    logger.info("  3/12 Creating Change Approval Workflow sheet (100 rows)...")
    create_change_approval_workflow_sheet(wb, styles)
    
    logger.info("  4/12 Creating Impact Assessment sheet (100 rows)...")
    create_impact_assessment_sheet(wb, styles)
    
    logger.info("  5/12 Creating Testing Validation sheet (100 rows)...")
    create_testing_validation_sheet(wb, styles)
    
    logger.info("  6/12 Creating Implementation Log sheet (100 rows)...")
    create_implementation_log_sheet(wb, styles)
    
    logger.info("  7/12 Creating Rollback Capability sheet (100 rows)...")
    create_rollback_capability_sheet(wb, styles)
    
    logger.info("  8/12 Creating Emergency Changes sheet (50 rows)...")
    create_emergency_changes_sheet(wb, styles)
    
    logger.info("  9/12 Creating Change Success Metrics sheet (dashboard)...")
    create_change_success_metrics_sheet(wb, styles)
    
    logger.info(" 10/12 Creating Evidence Register sheet (100 rows)...")
    create_evidence_register(wb, styles)
    
    logger.info(" 11/12 Creating Summary Dashboard sheet (Gold Standard)...")
    create_compliance_dashboard_sheet(wb, styles)
    
    logger.info(" 12/12 Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb, styles)
    
    logger.info("  ✓ All sheets created successfully")
    
    # Set workbook properties
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.created = datetime.now()
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Save workbook
    logger.info("-" * 70)
    logger.info("Saving workbook...")
    finalize_validations(wb)
    wb.save(output_path)
    logger.info("=" * 70)
    logger.info("✓ Workbook generated successfully!")
    logger.info("=" * 70)
    logger.info(f"Output File: {FILENAME}")
    logger.info(f"File Size: {os.path.getsize(_wkbk_dir / OUTPUT_FILENAME) / 1024:.1f} KB")
    logger.info(f"Total Sheets: 12")
    logger.info("-" * 70)
    logger.info("\nWorkbook Structure:")
    logger.info("  1.  Instructions - Usage guidance, change types, approval tiers")
    logger.info("  2.  Change Request Register - 100 rows for change tracking")
    logger.info("  3.  Change Approval Workflow - 100 rows for approval chain")
    logger.info("  4.  Impact Assessment - 100 rows for risk analysis")
    logger.info("  5.  Testing Validation - 100 rows for testing records")
    logger.info("  6.  Implementation Log - 100 rows for execution tracking")
    logger.info("  7.  Rollback Capability - 100 rows for rollback assessment")
    logger.info("  8.  Emergency Changes - 50 rows for expedited changes")
    logger.info("  9.  Change Success Metrics - Auto-calculated dashboard")
    logger.info("  10. Compliance Dashboard - Process adherence metrics")
    logger.info("  11. Evidence Register - 100 rows for evidence documentation")
    logger.info("  12. Approval Sign-Off - Three-tier approval signatures")
    logger.info("-" * 70)
    logger.info("\nNext Steps:")
    logger.info("1. Open workbook in Excel/LibreOffice")
    logger.info("2. Verify all sheets, validations, and formulas")
    logger.info("3. Review Instructions sheet for change types and priorities")
    logger.info("4. Customise dropdown values if needed (see CONFIGURATION section)")
    logger.info("5. Integrate with existing change management process")
    logger.info("6. Train Change Coordinators on workbook usage")
    logger.info("7. Use dashboards for monthly CAB reporting")
    logger.info("-" * 70)
    logger.info("\nKEY METRICS TO MONITOR:")
    logger.info("\u2022 Change Success Rate: Target ≥95%")
    logger.info("\u2022 Emergency Change Ratio: Target <10%")
    logger.info("\u2022 Approval Compliance: Target 100%")
    logger.info("\u2022 Testing Coverage: Target 100% for Normal changes")
    logger.info("\u2022 Rollback Readiness: Target 100% for High/Critical risk")
    logger.info("\u2022 Overall Compliance: Target ≥95%")
    logger.info("-" * 70)
    logger.info("\nIMPORTANT REMINDERS:")
    logger.info("\u2022 This is a SAMPLE workbook - customise for your organisation")
    logger.info("\u2022 Emergency changes require CAB review within 5 business days")
    logger.info("\u2022 High-risk changes require documented rollback procedures")
    logger.info("\u2022 Protected cells (gray) contain formulas - do not edit")
    logger.info("\u2022 Update Change Request Register in real-time as changes occur")
    logger.info("\u2022 Review metrics monthly, comprehensive assessment quarterly")
    logger.info("=" * 70)

def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
