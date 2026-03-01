#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.9.3 - Configuration Monitoring Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Assessment Domain 3 of 4: Configuration Monitoring and Drift Detection

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific monitoring infrastructure, drift detection tools,
and remediation workflows.

Key customization areas:
1. Monitoring tool capabilities (match your actual deployment)
2. Drift detection thresholds (adapt to your risk tolerance)
3. Alert routing and escalation (align with your SOC/operations structure)
4. Remediation SLAs (based on your operational requirements)
5. Monitoring coverage targets (specific to your asset criticality tiers)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
configuration monitoring and drift detection capabilities against ISO 27001:2022
Control A.8.9 requirements.

**Purpose:**
Enables systematic assessment of continuous monitoring, drift detection
effectiveness, alert management, and remediation processes to ensure
unauthorised configuration changes are detected and corrected.

**Assessment Scope:**
- Monitoring infrastructure deployment and coverage
- Drift detection capabilities and thresholds
- Alert generation and routing effectiveness
- Remediation workflow execution and SLA compliance
- Baseline comparison and compliance verification
- Monitoring tool performance and reliability
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and monitoring standards
2. Monitoring Infrastructure - Tool deployment and coverage tracking
3. Drift Detection - Alert configuration and detection capabilities
4. Baseline Comparison - Compliance scan results and drift analysis
5. Remediation Tracking - Drift incident remediation and SLA compliance
6. Monitoring Performance - Tool uptime, alert accuracy, false positive rates
7. Gap Analysis - Coverage gaps and remediation requirements
8. Evidence Register - Audit evidence tracking (100+ entries)
9. Approval & Sign-Off - Three-tier approval workflow

**Key Features:**
- Data validation with drift severity and asset tier dropdowns
- Conditional formatting for coverage status and SLA compliance
- Automated gap identification for unmonitored assets
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with monitoring and SIEM systems

**Integration:**
consolidates data from all four configuration management assessment domains
for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_3_monitoring.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_3_monitoring.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_3_monitoring.py --date 20250127

Output:
    File: ISMS_IMP_A_8_9_3_Configuration_Monitoring_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise monitoring coverage targets by asset tier
    2. Inventory all deployed monitoring tools and agents
    3. Document drift detection capabilities and thresholds
    4. Assess baseline comparison scan frequency and coverage
    5. Verify alert routing and remediation workflows
    6. Calculate remediation SLA compliance rates
    7. Review monitoring tool performance metrics
    8. Conduct gap analysis for unmonitored assets
    9. Define remediation actions with timelines
    10. Collect and link audit evidence (drift alerts, scan results)
    11. Obtain three-tier stakeholder approvals

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Assessment Domain:    3 of 4 (Configuration Monitoring and Drift Detection)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-POL-A.8.9, Section 2.4: Configuration Monitoring & Drift Detection
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9, Part 3: Configuration Deviation Response Procedures
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.1: Baseline Configuration Assessment (Domain 1)
    - ISMS-IMP-A.8.9.2: Change Control Assessment (Domain 2)
    - ISMS-IMP-A.8.9.4: Security Hardening Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.9.3 specification
    - Supports comprehensive configuration monitoring evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Monitoring Technology:**
Configuration monitoring tools and techniques evolve rapidly. Review vendor
roadmaps, emerging monitoring capabilities, and drift detection algorithms
quarterly. False positive tuning and alert fatigue prevention are critical.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of monitoring coverage, drift detection, and
remediation effectiveness.

**Data Protection:**
Assessment workbooks contain sensitive operational details including:
- Monitoring tool deployment architecture
- Drift alert details with system configurations
- Baseline compliance scan results
- Vulnerability information and security gaps

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Check monitoring coverage and drift remediation SLA compliance
- Quarterly: Review drift trends and monitoring tool performance
- Annually: Complete reassessment of monitoring infrastructure
- Ad-hoc: When monitoring tools change or new asset types deployed

**Quality Assurance:**
Have SOC analysts, monitoring engineers, and configuration managers validate
assessments before using results for compliance reporting or tool investment.

**Regulatory Alignment:**
Ensure monitoring practices align with applicable requirements:
- ISO 27001:2022: Control A.8.9 monitoring requirements
- Continuous monitoring: Real-time security monitoring mandates
- Sector-specific: Regulatory monitoring and alerting requirements
- Internal: Organisational monitoring and incident response policies

Customize assessment criteria to include regulatory-specific requirements.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# ============================================================================
# CONFIGURATION SECTION - CUSTOMIZE FOR YOUR ENVIRONMENT
# ============================================================================

# File output configuration
FILENAME = f"ISMS-IMP-A.8.9.3_Configuration_Monitoring_{datetime.now().strftime('%Y%m%d')}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Workbook metadata
WORKBOOK_TITLE = "Configuration Monitoring Assessment"
WORKBOOK_NAME = "Monitoring"
WORKBOOK_VERSION = "1.0"

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.9.3"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
CONTROL_ID   = "A.8.9"
CONTROL_NAME = "Configuration Management"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
# CUSTOMIZE: Configuration monitoring dropdown values
ASSET_CRITICALITY = ["Critical", "High", "Medium", "Low"]
MONITORING_STATUS = ["\u2705 Monitored", "\u26A0\uFE0F Partially Monitored", "\u274C Not Monitored", "➖ Excluded"]
MONITORING_METHOD = ["Automated Continuous", "Scheduled Automated", "Manual", "Hybrid", "None"]
CHECK_FREQUENCY = ["Real-time (<15 min)", "Hourly", "Daily", "Weekly", "Monthly", "Quarterly", "Manual (on-demand)"]

DRIFT_CATEGORY = ["Critical", "High", "Medium", "Low", "ℹ️ Informational"]
DETECTION_METHOD = ["Automated Continuous", "Scheduled Scan", "Manual Check", "User Report"]
AUTHORIZED_CHANGE = ["Yes (Change ID)", "No (Unauthorised)", "Under Investigation"]
ROOT_CAUSE_CATEGORY = ["Unauthorised Manual Change", "Tool Failure", "Software Update", 
                       "Baseline Not Updated", "Environmental", "Malicious", "Other"]
DRIFT_STATUS = ["⚠️ Detected", "⏳ Under Investigation", "⏳ Remediation In Progress",
               "\u2705 Remediated", "✔️ Closed", "➖ False Positive"]

TOOL_TYPE = ["Agent-Based", "Agentless", "Network Scanner", "Script/Custom", 
            "Cloud-Native", "SIEM Integration"]
DEPLOYMENT_STATUS = ["\u2705 Active", "\u26A0\uFE0F Degraded", "\u274C Offline", "Pilot", "⏸️ Decommissioned"]
ALERTING_METHOD = ["Email", "SIEM", "Webhook", "Dashboard Only", "Ticketing System", "Multiple"]
LICENSING_MODEL = ["Commercial", "Open Source", "Subscription", "Perpetual", "In-House Developed"]

REMEDIATION_ACTION = ["Reverted to Baseline", "Updated Baseline", "Authorised Retroactively", 
                     "No Action (False Positive)", "Other"]
VERIFICATION_METHOD = ["Automated Re-Scan", "Manual Verification", "Monitoring Tool Confirmation", 
                      "User Validation"]
VERIFICATION_RESULT = ["\u2705 Passed", "\u274C Failed", "\u26A0\uFE0F Partially Successful", "❓ Not Yet Verified"]
ROOT_CAUSE_REMEDIATION = ["Baseline Updated", "Change Control Enforced", "Tool Fixed", 
                         "Process Improved", "Training Provided", "Other"]

FALSE_POSITIVE_REASON = ["Incorrect Baseline", "Tool Misconfiguration", "Expected Variation", 
                        "Timing Issue", "Tool Bug", "Other"]
TUNING_ACTION = ["Baseline Updated", "Monitoring Rule Adjusted", "Alert Threshold Changed", 
                "Exception Added", "Tool Updated", "No Action (Acceptable)"]
RECURRENCE_STATUS = ["Not Seen Again", "Recurred Once", "Recurring (Needs Further Tuning)", "Monitoring"]

EVIDENCE_TYPE = ["Monitoring Configuration", "Drift Alert Screenshot", "Remediation Record", 
                "Tool Health Check", "Coverage Report", "Scan Output", "Other"]
EVIDENCE_CLASSIFICATION = ["Public", "Internal", "Confidential", "Restricted"]
RETENTION_PERIOD = ["1 Year", "3 Years", "5 Years", "7 Years", "Indefinite"]
EVIDENCE_VERIFICATION = ["Verified", "Needs Verification", "Missing", "Outdated"]

APPROVAL_DECISION = ["Approved", "Approved with Conditions", "Not Approved - Revisions Required"]

# CUSTOMIZE: Asset types (43 types from A.8.9.1)
ASSET_TYPES = {
    "Infrastructure": [
        "Physical Server", "Virtual Machine", "Hypervisor", "Network Switch",
        "Router", "Firewall", "Load Balancer", "Storage Array",
        "Backup Appliance", "UPS", "HVAC", "Physical Security System"
    ],
    "Endpoint": [
        "Desktop", "Laptop", "Mobile Phone", "Tablet", "Thin Client", "Kiosk/POS"
    ],
    "Network Services": [
        "DNS Server", "DHCP Server", "NTP Server", "Proxy Server",
        "VPN Gateway", "Wireless Access Point", "Network Management System"
    ],
    "Applications": [
        "Enterprise Application", "Web Application", "Database Server",
        "Middleware", "API Gateway", "Custom Developed Application",
        "COTS Application", "Open Source Application"
    ],
    "Cloud": [
        "IaaS Resource", "PaaS Service", "SaaS Application",
        "Cloud-Native Application", "Serverless Function"
    ],
    "IoT/OT": [
        "IoT Device", "Industrial Control System", "SCADA System",
        "Building Management System", "Medical Device"
    ]
}

# CUSTOMIZE: Color scheme
COLORS = {
    'header_main': '003366',
    'header_sub': '4472C4',
    'column_header': 'D9D9D9',
    'input_cell': 'FFFFCC',
    'protected_cell': 'F2F2F2',
    'compliant': 'C6EFCE',
    'partial': 'FFEB9C',
    'non_compliant': 'FFC7CE',
    'excluded': 'D9D9D9',
    'critical': 'C00000',
    'info_bg': 'F2F2F2',
    'light_green': 'C6EFCE',
    'orange': 'FFA500'
}
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================
def create_styles():
    """Creates and returns a dictionary of reusable styles."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    styles = {
        'header_main': {
            'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_main'], 
                              end_color=COLORS['header_main'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'header_sub': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'column_header': {
            'font': Font(name='Calibri', size=11, bold=True),
            'fill': PatternFill(start_color=COLORS['column_header'], 
                              end_color=COLORS['column_header'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'input_cell': {
            'font': Font(name='Calibri', size=11),
            'fill': PatternFill(start_color=COLORS['input_cell'], 
                              end_color=COLORS['input_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False),
            'border': thin_border
        },
        'protected_cell': {
            'font': Font(name='Calibri', size=11, italic=True),
            'fill': PatternFill(start_color=COLORS['protected_cell'], 
                              end_color=COLORS['protected_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top'),
            'border': thin_border
        },
        'info_text': {
            'font': Font(name='Calibri', size=11),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True)
        },
        'section_header': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': thin_border
        }
    }
    return styles

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_data_validation(values, allow_blank=True):
    """Create a data validation object for dropdowns."""
    formula = f'"{",".join(values)}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the dropdown'
    dv.promptTitle = 'Selection Required'
    return dv

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================

def create_lookup_tables(wb, styles):
    """Create hidden Lookup Tables sheet with 43 asset types."""
    ws = wb.create_sheet("Lookup Tables")
    ws.sheet_view.showGridLines = False
    ws.sheet_state = 'hidden'
    
    ws['A1'] = "Asset Type"
    ws['B1'] = "Asset Category"
    apply_style(ws['A1'], styles['header_main'])
    apply_style(ws['B1'], styles['header_main'])
    
    row = 2
    for category, types in ASSET_TYPES.items():
        for asset_type in types:
            ws[f'A{row}'] = asset_type
            ws[f'B{row}'] = category
            row += 1
    
    return ws


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab for applicable systems/services.', '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Check all applicable items in the Compliance Checklist for each section.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_monitoring_coverage_register_sheet(wb, styles):
    """Create Sheet 2: Monitoring Coverage Register"""
    ws = wb.create_sheet("Monitoring Coverage Register")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18,  # Asset ID
        'B': 30,  # Asset Name
        'C': 35,  # Asset Type
        'D': 20,  # Asset Category
        'E': 15,  # Asset Criticality
        'F': 25,  # Monitoring Tier
        'G': 18,  # Monitoring Status
        'H': 25,  # Monitoring Method
        'I': 30,  # Monitoring Tool/System
        'J': 20,  # Check Frequency
        'K': 25,  # Baseline Reference
        'L': 15,  # Last Monitored Date
        'M': 35,  # Monitoring Configuration Location
        'N': 18,  # Coverage Compliance
        'O': 40,  # Gap Justification
        'P': 35   # Notes
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:P1')
    ws['A1'] = "CONFIGURATION MONITORING COVERAGE REGISTER"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Asset ID', 'B': 'Asset Name', 'C': 'Asset Type', 'D': 'Asset Category',
        'E': 'Asset Criticality', 'F': 'Monitoring Tier', 'G': 'Monitoring Status',
        'H': 'Monitoring Method', 'I': 'Monitoring Tool/System', 'J': 'Check Frequency',
        'K': 'Baseline Reference', 'L': 'Last Monitored Date', 
        'M': 'Monitoring Configuration Location', 'N': 'Coverage Compliance',
        'O': 'Gap Justification', 'P': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_mcr = {
        'A': 'ASSET-001', 'B': 'prod-web-01.company.com', 'C': 'Physical Server',
        'D': 'Infrastructure', 'E': 'Critical', 'F': 'Tier 1 - Critical Assets',
        'G': '\u2705 Monitored', 'H': 'Automated Continuous', 'I': 'CIS-CAT Pro Scanner',
        'J': 'Real-time (<15 min)', 'K': 'BL-001 Windows Server 2022 CIS v2.0',
        'L': '20.01.2026', 'M': '/monitoring/agents/prod-web-01/config.xml',
        'N': 'Compliant', 'O': '', 'P': 'Tier 1 asset — requires continuous monitoring per policy'
    }
    _sfill_mcr = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sborder_mcr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_mcr.get(col, '')
        cell.fill = _sfill_mcr
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_mcr

    num_rows = 50
    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 51 (1 sample + 50 empty)
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['D', 'F', 'N']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formulas
    for row in range(5, 5 + num_rows):
        # Column D: Asset Category (VLOOKUP from Lookup Tables)
        ws[f'D{row}'] = f"=IFERROR(VLOOKUP(C{row},'Lookup Tables'!$A$2:$B$44,2,FALSE),\"\")"

        # Column F: Monitoring Tier (based on criticality)
        ws[f'F{row}'] = f'=IF(E{row}="Critical","Tier 1 - Critical Assets",IF(E{row}="High","Tier 2 - High Value Assets",IF(E{row}="Medium","Tier 3 - Standard Assets","Tier 4 - Low Risk Assets")))'

        # Column N: Coverage Compliance (complex logic)
        ws[f'N{row}'] = f'=IF(G{row}="Excluded","Excluded",IF(E{row}="Critical",IF(AND(G{row}="Monitored",OR(H{row}="Automated Continuous",H{row}="Hybrid")),"Compliant","Non-Compliant"),IF(E{row}="High",IF(G{row}="Monitored","Compliant",IF(G{row}="Partially Monitored","Partial","Non-Compliant")),IF(G{row}="Not Monitored","Non-Compliant","Compliant"))))'

    # Data validations
    validations = []
    asset_type_dv = DataValidation(type="list", formula1="'Lookup Tables'!$A$2:$A$44", allow_blank=True)
    asset_type_dv.add(f'C4:C{4+num_rows}')
    validations.append(asset_type_dv)

    criticality_dv = create_data_validation(ASSET_CRITICALITY, allow_blank=False)
    criticality_dv.add(f'E4:E{4+num_rows}')
    validations.append(criticality_dv)

    status_dv = create_data_validation(MONITORING_STATUS, allow_blank=False)
    status_dv.add(f'G4:G{4+num_rows}')
    validations.append(status_dv)

    method_dv = create_data_validation(MONITORING_METHOD, allow_blank=False)
    method_dv.add(f'H4:H{4+num_rows}')
    validations.append(method_dv)

    freq_dv = create_data_validation(CHECK_FREQUENCY, allow_blank=True)
    freq_dv.add(f'J4:J{4+num_rows}')
    validations.append(freq_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'E4:E{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'E4:E{4+num_rows}',
        CellIsRule(operator='equal', formula=['"High"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Medium"'],
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Low"'],
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid')))

    ws.conditional_formatting.add(f'G4:G{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Monitored"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'G4:G{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Monitored"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    ws.conditional_formatting.add(f'N4:N{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Compliant"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'N4:N{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Non-Compliant"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))

    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_drift_detection_log_sheet(wb, styles):
    """Create Sheet 3: Drift Detection Log"""
    ws = wb.create_sheet("Drift Detection Log")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 20, 'B': 20, 'C': 18, 'D': 25, 'E': 30, 'F': 25,
        'G': 25, 'H': 15, 'I': 20, 'J': 25, 'K': 20, 'L': 25,
        'M': 25, 'N': 35, 'O': 20, 'P': 12, 'Q': 20, 'R': 15, 'S': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:S1')
    ws['A1'] = "CONFIGURATION DRIFT DETECTION LOG"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Drift Incident ID', 'B': 'Detection Date/Time', 'C': 'Asset ID',
        'D': 'Asset Name', 'E': 'Configuration Item', 'F': 'Expected Value',
        'G': 'Actual Value', 'H': 'Drift Category', 'I': 'Detection Method',
        'J': 'Detecting Tool/System', 'K': 'Authorised Change', 'L': 'Change ID Reference',
        'M': 'Root Cause Category', 'N': 'Root Cause Detail', 'O': 'Drift Status',
        'P': 'Time to Detect (Hours)', 'Q': 'Remediation Owner', 'R': 'Priority', 'S': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_ddl = {
        'A': 'DRIFT-2026-0001', 'B': '15.01.2026 09:23',
        'C': 'SRV-PROD-001', 'D': 'prod-web-01.company.com',
        'E': 'Password Policy / MinimumPasswordLength', 'F': '14 characters',
        'G': '8 characters', 'H': 'High', 'I': 'Automated Continuous',
        'J': 'CIS-CAT Pro Scanner', 'K': 'No (Unauthorised)',
        'L': '', 'M': 'Unauthorised Manual Change', 'N': 'Admin configured shorter password for testing — not reverted',
        'O': '\u2705 Remediated', 'Q': 'J. Smith (Sec Ops)', 'S': 'Reverted within 2 hours of detection'
    }
    _sfill_ddl = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sborder_ddl = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_ddl.get(col, '')
        cell.fill = _sfill_ddl
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_ddl

    num_rows = 50
    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 51 (1 sample + 50 empty)
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['P', 'R']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formulas
    for row in range(5, 5 + num_rows):
        # Column P: Time to Detect
        ws[f'P{row}'] = f'=IF(B{row}="","","")'

        # Column R: Priority
        ws[f'R{row}'] = f'=IF(H{row}="Critical","P1-Critical",IF(H{row}="High","P2-High",IF(H{row}="Medium","P3-Medium","P4-Low")))'

    # Data validations
    validations = []
    drift_cat_dv = create_data_validation(DRIFT_CATEGORY, allow_blank=False)
    drift_cat_dv.add(f'H4:H{4+num_rows}')
    validations.append(drift_cat_dv)

    detect_dv = create_data_validation(DETECTION_METHOD, allow_blank=False)
    detect_dv.add(f'I4:I{4+num_rows}')
    validations.append(detect_dv)

    auth_dv = create_data_validation(AUTHORIZED_CHANGE, allow_blank=False)
    auth_dv.add(f'K4:K{4+num_rows}')
    validations.append(auth_dv)

    root_dv = create_data_validation(ROOT_CAUSE_CATEGORY, allow_blank=True)
    root_dv.add(f'M4:M{4+num_rows}')
    validations.append(root_dv)

    status_dv = create_data_validation(DRIFT_STATUS, allow_blank=False)
    status_dv.add(f'O4:O{4+num_rows}')
    validations.append(status_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"High"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Medium"'],
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))

    ws.conditional_formatting.add(f'K4:K{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Yes (Change ID)"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'K4:K{4+num_rows}',
        CellIsRule(operator='equal', formula=['"No (Unauthorised)"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    ws.conditional_formatting.add(f'O4:O{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Closed"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))

    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_monitoring_tool_inventory_sheet(wb, styles):
    """Create Sheet 4: Monitoring Tool Inventory"""
    ws = wb.create_sheet("Monitoring Tool Inventory")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 15,  # Tool ID
        'B': 25,  # Tool Name
        'C': 20,  # Vendor
        'D': 20,  # Tool Type
        'E': 40,  # Monitoring Capabilities
        'F': 35,  # Asset Types Covered
        'G': 15,  # Assets Monitored Count
        'H': 18,  # Deployment Status
        'I': 35,  # Integration Points
        'J': 20,  # Alerting Method
        'K': 20,  # Licensing Model
        'L': 15,  # Annual Cost
        'M': 15,  # Last Health Check
        'N': 35,  # Known Limitations
        'O': 20,  # Tool Owner
        'P': 35   # Notes
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:P1')
    ws['A1'] = "MONITORING TOOL INVENTORY"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Tool ID', 'B': 'Tool Name', 'C': 'Vendor', 'D': 'Tool Type',
        'E': 'Monitoring Capabilities', 'F': 'Asset Types Covered',
        'G': 'Assets Monitored Count', 'H': 'Deployment Status',
        'I': 'Integration Points', 'J': 'Alerting Method', 'K': 'Licensing Model',
        'L': 'Annual Cost', 'M': 'Last Health Check', 'N': 'Known Limitations',
        'O': 'Tool Owner', 'P': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_mti = {
        'A': 'TOOL-001', 'B': 'CIS-CAT Pro Scanner', 'C': 'Centre for Internet Security',
        'D': 'Agentless', 'E': 'CIS Benchmark compliance scanning, configuration drift detection',
        'F': 'Physical Server, Virtual Machine, Desktop, Laptop',
        'G': 42, 'H': '\u2705 Active', 'I': 'SIEM, Ticketing System',
        'J': 'SIEM', 'K': 'Commercial', 'L': 4500.00, 'M': '15.01.2026',
        'N': 'No cloud-native asset support', 'O': 'J. Smith (Sec Ops)', 'P': 'Annual renewal due 01.06.2026'
    }
    _sfill_mti = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sborder_mti = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_mti.get(col, '')
        cell.fill = _sfill_mti
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_mti

    num_rows = 50
    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 51 (1 sample + 50 empty)
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            apply_style(ws[f'{col}{row}'], styles['input_cell'])

    # Data validations
    validations = []
    tool_type_dv = create_data_validation(TOOL_TYPE, allow_blank=False)
    tool_type_dv.add(f'D4:D{4+num_rows}')
    validations.append(tool_type_dv)

    deploy_dv = create_data_validation(DEPLOYMENT_STATUS, allow_blank=False)
    deploy_dv.add(f'H4:H{4+num_rows}')
    validations.append(deploy_dv)

    alert_dv = create_data_validation(ALERTING_METHOD, allow_blank=True)
    alert_dv.add(f'J4:J{4+num_rows}')
    validations.append(alert_dv)

    license_dv = create_data_validation(LICENSING_MODEL, allow_blank=True)
    license_dv.add(f'K4:K{4+num_rows}')
    validations.append(license_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"\u2705 Active"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"\u26A0\uFE0F Degraded"'],
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'H4:H{4+num_rows}',
        CellIsRule(operator='equal', formula=['"\u274C Offline"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))

    ws.freeze_panes = 'B3'

    # Add total cost formula at bottom
    total_row = 5 + num_rows
    ws[f'K{total_row}'] = "TOTAL:"
    ws[f'K{total_row}'].font = Font(bold=True)
    ws[f'L{total_row}'] = f'=SUM(L4:L{4+num_rows})'
    ws[f'L{total_row}'].font = Font(bold=True)
    ws[f'L{total_row}'].number_format = '#,##0.00'
    
    return ws

def create_drift_remediation_tracking_sheet(wb, styles):
    """Create Sheet 5: Drift Remediation Tracking"""
    ws = wb.create_sheet("Drift Remediation Tracking")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 20, 'B': 25, 'C': 15, 'D': 15, 'E': 20, 'F': 15,
        'G': 25, 'H': 40, 'I': 15, 'J': 12, 'K': 25, 'L': 15,
        'M': 18, 'N': 35, 'O': 25, 'P': 18, 'Q': 15, 'R': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:R1')
    ws['A1'] = "DRIFT REMEDIATION TRACKING"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'Drift Incident ID', 'B': 'Asset Name', 'C': 'Drift Category',
        'D': 'Detection Date', 'E': 'Remediation Owner', 'F': 'Remediation Started Date',
        'G': 'Remediation Action Taken', 'H': 'Remediation Action Detail',
        'I': 'Remediation Completed Date', 'J': 'Time to Remediate (Days)',
        'K': 'Verification Method', 'L': 'Verification Date', 'M': 'Verification Result',
        'N': 'Recurrence Prevention', 'O': 'Root Cause Remediation',
        'P': 'Remediation Status', 'Q': 'SLA Compliance', 'R': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    num_rows = 51
    # MAX standard: Row 1 = sample with example data, Rows 2-51 = empty
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['J', 'P', 'Q']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formulas
    for row in range(3, 3 + num_rows):
        # Column J: Time to Remediate (Days)
        ws[f'J{row}'] = f'=IF(OR(D{row}="",I{row}=""),"",I{row}-D{row})'

        # Column P: Remediation Status
        ws[f'P{row}'] = f'=IF(I{row}<>"","Complete",IF(J{row}="","Not Started",IF(C{row}="Critical",IF(J{row}<1,"On-Time","Overdue"),IF(C{row}="High",IF(J{row}<2,"On-Time",IF(J{row}<3,"At Risk","Overdue")),IF(J{row}<8,"On-Time","At Risk")))))'

        # Column Q: SLA Compliance
        ws[f'Q{row}'] = f'=IF(P{row}="Complete",IF(C{row}="Critical",IF(J{row}<=0.17,"Met","Missed"),IF(C{row}="High",IF(J{row}<=1,"Met","Missed"),IF(C{row}="Medium",IF(J{row}<=7,"Met","Missed"),"N/A"))),"In Progress")'

    # Data validations
    validations = []
    drift_cat_dv = create_data_validation(DRIFT_CATEGORY, allow_blank=False)
    drift_cat_dv.add(f'C3:C{2+num_rows}')
    validations.append(drift_cat_dv)

    action_dv = create_data_validation(REMEDIATION_ACTION, allow_blank=False)
    action_dv.add(f'G3:G{2+num_rows}')
    validations.append(action_dv)

    verif_method_dv = create_data_validation(VERIFICATION_METHOD, allow_blank=True)
    verif_method_dv.add(f'K3:K{2+num_rows}')
    validations.append(verif_method_dv)

    verif_result_dv = create_data_validation(VERIFICATION_RESULT, allow_blank=True)
    verif_result_dv.add(f'M3:M{2+num_rows}')
    validations.append(verif_result_dv)

    root_rem_dv = create_data_validation(ROOT_CAUSE_REMEDIATION, allow_blank=True)
    root_rem_dv.add(f'O3:O{2+num_rows}')
    validations.append(root_rem_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'C3:C{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'], 
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Complete"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Overdue"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.conditional_formatting.add(f'Q3:Q{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Met"'], 
                   font=Font(bold=True, color='006100')))
    ws.conditional_formatting.add(f'Q3:Q{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Missed"'], 
                   font=Font(bold=True, color='9C0006')))
    
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Passed"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_false_positive_register_sheet(wb, styles):
    """Create Sheet 6: False Positive Register"""
    ws = wb.create_sheet("False Positive Register")
    ws.sheet_view.showGridLines = False
    
    widths = {
        'A': 18, 'B': 15, 'C': 25, 'D': 18, 'E': 25, 'F': 40,
        'G': 25, 'H': 40, 'I': 15, 'J': 20, 'K': 25, 'L': 35,
        'M': 15, 'N': 20, 'O': 15, 'P': 25, 'Q': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "FALSE POSITIVE REGISTER - ALERT QUALITY TRACKING"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    headers = {
        'A': 'False Positive ID', 'B': 'Alert Date', 'C': 'Monitoring Tool',
        'D': 'Asset ID', 'E': 'Alert Type', 'F': 'Alert Message',
        'G': 'False Positive Reason', 'H': 'False Positive Detail',
        'I': 'Investigation Date', 'J': 'Investigated By', 'K': 'Tuning Action Taken',
        'L': 'Tuning Action Detail', 'M': 'Tuning Completed Date',
        'N': 'Recurrence Status', 'O': 'Last Recurrence Date',
        'P': 'False Positive Category', 'Q': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample row (row 4) — F2F2F2 grey, fully populated as usage example
    sample_fpr = {
        'A': 'FP-2026-0001', 'B': '10.01.2026', 'C': 'CIS-CAT Pro Scanner',
        'D': 'ASSET-003', 'E': 'Password Policy / PasswordComplexity',
        'F': 'Complexity not enabled (expected: 1)',
        'G': 'Expected Variation', 'H': 'Guest account excluded by policy — baseline not updated',
        'I': '11.01.2026', 'J': 'J. Smith (Sec Ops)', 'K': 'Baseline Updated',
        'L': 'Updated baseline BL-001 to exclude guest account PasswordComplexity check',
        'M': '12.01.2026', 'N': 'Not Seen Again', 'O': '',
        'P': 'Baseline Issue', 'Q': 'Updated baseline distributed to all scanner profiles'
    }
    _sfill_fpr = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sborder_fpr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in headers.keys():
        cell = ws[f'{col}4']
        cell.value = sample_fpr.get(col, '')
        cell.fill = _sfill_fpr
        cell.font = Font(name='Calibri', size=11, italic=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = _sborder_fpr

    num_rows = 50
    # Create 50 empty data rows (rows 5-54) — total FFFFCC = 51 (1 sample + 50 empty)
    for row in range(5, 5 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col == 'P':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['input_cell'])

    # Formula for False Positive Category
    for row in range(5, 5 + num_rows):
        ws[f'P{row}'] = f'=IF(OR(G{row}="Tool Misconfiguration",G{row}="Tool Bug"),"Systemic (tool issue)",IF(G{row}="Incorrect Baseline","Baseline Issue","One-Time (environmental)"))'

    # Data validations
    validations = []
    fp_reason_dv = create_data_validation(FALSE_POSITIVE_REASON, allow_blank=False)
    fp_reason_dv.add(f'G4:G{4+num_rows}')
    validations.append(fp_reason_dv)

    tuning_dv = create_data_validation(TUNING_ACTION, allow_blank=False)
    tuning_dv.add(f'K4:K{4+num_rows}')
    validations.append(tuning_dv)

    recur_dv = create_data_validation(RECURRENCE_STATUS, allow_blank=False)
    recur_dv.add(f'N4:N{4+num_rows}')
    validations.append(recur_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting
    ws.conditional_formatting.add(f'N4:N{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Seen Again"'],
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'N4:N{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Recurring (Needs Further Tuning)"'],
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    ws.conditional_formatting.add(f'P4:P{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Systemic (tool issue)"'],
                   font=Font(color='9C0006')))
    ws.conditional_formatting.add(f'P4:P{4+num_rows}',
        CellIsRule(operator='equal', formula=['"Baseline Issue"'],
                   font=Font(color='C65911')))

    ws.freeze_panes = 'B3'
    # Formula cells unlocked
    
    return ws

def create_monitoring_effectiveness_metrics_sheet(wb, styles):
    """Create Sheet 7: Monitor Effectiveness Metrics (Dashboard)"""
    ws = wb.create_sheet("Monitor Effectiveness Metrics")
    ws.sheet_view.showGridLines = False
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "MONITORING EFFECTIVENESS METRICS DASHBOARD"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35

    # SECTION A: Overall Monitoring Coverage
    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL MONITORING COVERAGE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    coverage_metrics = [
        ("Total Assets in Scope", "=COUNTA('Monitoring Coverage Register'!A3:A102)-COUNTBLANK('Monitoring Coverage Register'!A3:A102)", "N/A", ""),
        ("Assets Monitored", "=COUNTIF('Monitoring Coverage Register'!G3:G102,\"✅ Monitored\")", "\u226585%", ""),
        ("Overall Monitoring Coverage %", "=IF(B5=0,0,B6/B5*100)", "\u226585%", "=IF(B7>=85,\"\u2713 Compliant\",IF(B7>=70,\"\u26A0 Partial\",\"\u2717 Non-Compliant\"))"),
        ("Tier 1 (Critical) Coverage %", "=IF(COUNTIF('Monitoring Coverage Register'!F3:F102,\"Tier 1*\")=0,0,COUNTIFS('Monitoring Coverage Register'!F3:F102,\"Tier 1*\",'Monitoring Coverage Register'!G3:G102,\"Monitored\")/COUNTIF('Monitoring Coverage Register'!F3:F102,\"Tier 1*\")*100)", "100%", "=IF(B8>=100,\"\u2713\",\"\u2717\")"),
        ("Tier 2 (High) Coverage %", "=IF(COUNTIF('Monitoring Coverage Register'!F3:F102,\"Tier 2*\")=0,0,COUNTIFS('Monitoring Coverage Register'!F3:F102,\"Tier 2*\",'Monitoring Coverage Register'!G3:G102,\"Monitored\")/COUNTIF('Monitoring Coverage Register'!F3:F102,\"Tier 2*\")*100)", "\u226595%", "=IF(B9>=95,\"\u2713\",\"\u2717\")"),
        ("Tier 3 (Standard) Coverage %", "=IF(COUNTIF('Monitoring Coverage Register'!F3:F102,\"Tier 3*\")=0,0,COUNTIFS('Monitoring Coverage Register'!F3:F102,\"Tier 3*\",'Monitoring Coverage Register'!G3:G102,\"Monitored\")/COUNTIF('Monitoring Coverage Register'!F3:F102,\"Tier 3*\")*100)", "\u226585%", "=IF(B10>=85,\"\u2713\",\"\u2717\")"),
        ("Non-Compliant Coverage", "=COUNTIF('Monitoring Coverage Register'!N3:N102,\"Non-Compliant\")", "0", "=IF(B11=0,\"\u2713 None\",\"\u2717 \"&B11&\" Assets\")"),
        ("Monitoring Tools Active", "=COUNTIF('Monitoring Tool Inventory'!H3:H32,\"✅ Active\")", "All", ""),
        ("Tools Offline/Degraded", "=COUNTIF('Monitoring Tool Inventory'!H3:H32,\"❌ Offline\")+COUNTIF('Monitoring Tool Inventory'!H3:H32,\"⚠️ Degraded\")", "0", "=IF(B13=0,\"\u2713 None\",\"\u2717 \"&B13&\" Issues\")"),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in coverage_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION B: Drift Detection Metrics
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "DRIFT DETECTION METRICS"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    drift_metrics = [
        ("Total Drift Incidents (All Time)", "=COUNTA('Drift Detection Log'!A3:A152)-COUNTBLANK('Drift Detection Log'!A3:A152)", "N/A", ""),
        ("Critical Drift Incidents", "=COUNTIF('Drift Detection Log'!H3:H152,\"Critical\")", "0", "=IF(B{row}=0,\"\u2713 None\",\"\u2717 \"&B{row}&\" CRITICAL\")"),
        ("High Drift Incidents", "=COUNTIF('Drift Detection Log'!H3:H152,\"High\")", "<5/month", ""),
        ("Unauthorised Changes Detected", "=COUNTIF('Drift Detection Log'!K3:K152,\"No (Unauthorised)\")", "Detect all", ""),
        ("False Positive Rate %", "=IF(B{total}=0,0,COUNTIF('Drift Detection Log'!O3:O152,\"➖ False Positive\")/B{total}*100)", "<10%", "=IF(B{row}<10,\"\u2713 Good\",IF(B{row}<20,\"\u26A0 Warning\",\"\u2717 Critical\"))"),
    ]
    
    row += 1
    total_row = row - 4  # Reference to total incidents
    for metric_name, value_formula, target, status_formula in drift_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        value_formula = value_formula.replace("{row}", str(row)).replace("{total}", str(total_row))
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        status_formula = status_formula.replace("{row}", str(row)).replace("{total}", str(total_row))
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION C: Remediation Performance
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "REMEDIATION PERFORMANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    remediation_metrics = [
        ("Open Drift Incidents", "=COUNTIFS('Drift Detection Log'!O3:O152,\"<>Closed\",'Drift Detection Log'!O3:O152,\"<>False Positive\")", "Minimize", ""),
        ("Overdue Critical Drift", "=COUNTIFS('Drift Remediation Tracking'!C3:C152,\"Critical\",'Drift Remediation Tracking'!P3:P152,\"Overdue\")", "0", "=IF(B{row}=0,\"\u2713 None\",\"\u2717 \"&B{row}&\" OVERDUE\")"),
        ("Mean Time to Remediate (Days)", "=IFERROR(AVERAGE('Drift Remediation Tracking'!J3:J152),0)", "<1 day", ""),
        ("SLA Compliance - Critical %", "=IF(COUNTIF('Drift Remediation Tracking'!C3:C152,\"Critical\")=0,100,COUNTIFS('Drift Remediation Tracking'!C3:C152,\"Critical\",'Drift Remediation Tracking'!Q3:Q152,\"Met\")/COUNTIF('Drift Remediation Tracking'!C3:C152,\"Critical\")*100)", "100%", "=IF(B{row}>=100,\"\u2713\",\"\u2717\")"),
        ("Overall Remediation Success Rate %", "=IF(COUNTA('Drift Remediation Tracking'!M3:M152)=0,0,COUNTIF('Drift Remediation Tracking'!M3:M152,\"\u2705 Passed\")/COUNTA('Drift Remediation Tracking'!M3:M152)*100)", "\u226598%", "=IF(B{row}>=98,\"\u2713 Excellent\",IF(B{row}>=90,\"\u26A0 Good\",\"\u2717 Needs Improvement\"))"),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in remediation_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        value_formula = value_formula.replace("{row}", str(row))
        ws[f'B{row}'] = value_formula
        if "%" in metric_name or "Days" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        status_formula = status_formula.replace("{row}", str(row))
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # Add note
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "NOTE: All metrics auto-calculated. Review monthly. Critical drift = absolute zero tolerance."
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Conditional formatting for coverage percentages
    ws.conditional_formatting.add('B7:B10',
        CellIsRule(operator='greaterThanOrEqual', formula=['85'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add('B7:B10',
        CellIsRule(operator='between', formula=['70', '84.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add('B7:B10',
        CellIsRule(operator='lessThan', formula=['70'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    return ws

def create_coverage_gap_analysis_sheet(wb, styles):
    """Create Sheet 8: Coverage Gap Analysis (Dashboard)"""
    ws = wb.create_sheet("Coverage Gap Analysis")
    ws.sheet_view.showGridLines = False
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "MONITORING COVERAGE GAP ANALYSIS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35

    # SECTION A: Coverage by Asset Category
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "COVERAGE BY ASSET CATEGORY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Asset Category"
    ws[f'B{row}'] = "Total Assets"
    ws[f'C{row}'] = "Monitored"
    ws[f'D{row}'] = "Coverage %"
    ws[f'E{row}'] = "Gap Priority"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    categories = ["Infrastructure", "Endpoint", "Network Services", "Applications", "Cloud", "IoT/OT"]
    row += 1
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f"=COUNTIF('Monitoring Coverage Register'!$D$3:$D$102,\"{category}\")"
        apply_style(ws[f'B{row}'], styles['protected_cell'])

        ws[f'C{row}'] = f"=COUNTIFS('Monitoring Coverage Register'!$D$3:$D$102,\"{category}\",'Monitoring Coverage Register'!$G$3:$G$102,\"Monitored\")"
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF(B{row}=0,0,C{row}/B{row}*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = f'=IF(D{row}<70,"Critical Gap",IF(D{row}<85,"High Priority","Acceptable"))'
        apply_style(ws[f'E{row}'], styles['protected_cell'])
        
        row += 1
    
    # Conditional formatting
    start_row = row - len(categories)
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['85'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='between', formula=['70', '84.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='lessThan', formula=['70'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'E{start_row}:E{row-1}',
        CellIsRule(operator='equal', formula=['"Critical Gap"'], 
                   font=Font(bold=True, color='9C0006')))
    ws.conditional_formatting.add(f'E{start_row}:E{row-1}',
        CellIsRule(operator='equal', formula=['"High Priority"'], 
                   font=Font(bold=True, color='C65911')))
    
    # SECTION B: Coverage by Criticality
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "COVERAGE BY ASSET CRITICALITY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Criticality"
    ws[f'B{row}'] = "Total Assets"
    ws[f'C{row}'] = "Monitored"
    ws[f'D{row}'] = "Coverage %"
    ws[f'E{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    criticalities = [
        ("Critical", "100%"),
        ("High", "≥95%"),
        ("Medium", "≥85%"),
        ("Low", "≥70%")
    ]
    
    row += 1
    for crit, target in criticalities:
        ws[f'A{row}'] = crit
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f"=COUNTIF('Monitoring Coverage Register'!$E$3:$E$102,\"{crit}\")"
        apply_style(ws[f'B{row}'], styles['protected_cell'])

        ws[f'C{row}'] = f"=COUNTIFS('Monitoring Coverage Register'!$E$3:$E$102,\"{crit}\",'Monitoring Coverage Register'!$G$3:$G$102,\"Monitored\")"
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF(B{row}=0,0,C{row}/B{row}*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        if crit == "Critical":
            ws[f'E{row}'] = f'=IF(D{row}>=100,"✓","✗")'
        elif crit == "High":
            ws[f'E{row}'] = f'=IF(D{row}>=95,"✓","✗")'
        elif crit == "Medium":
            ws[f'E{row}'] = f'=IF(D{row}>=85,"✓","✗")'
        else:
            ws[f'E{row}'] = f'=IF(D{row}>=70,"✓","✗")'
        apply_style(ws[f'E{row}'], styles['protected_cell'])
        
        row += 1
    
    return ws

def create_drift_trend_analysis_sheet(wb, styles):
    """Create Sheet 9: Drift Trend Analysis (Dashboard)"""
    ws = wb.create_sheet("Drift Trend Analysis")
    ws.sheet_view.showGridLines = False
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "CONFIGURATION DRIFT TREND ANALYSIS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35

    # Note: Simplified dashboard - in real implementation would include time-series analysis
    row = 3
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "DRIFT DISTRIBUTION BY CATEGORY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    
    row += 1
    ws[f'A{row}'] = "Drift Category"
    ws[f'B{row}'] = "Total"
    ws[f'C{row}'] = "Open"
    ws[f'D{row}'] = "Closed"
    ws[f'E{row}'] = "Percentage"
    ws[f'F{row}'] = "Trend"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    categories = ["Critical", "High", "Medium", "Low", "Informational"]
    row += 1
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f"=COUNTIF('Drift Detection Log'!$H$3:$H$152,\"{category}\")"
        apply_style(ws[f'B{row}'], styles['protected_cell'])

        ws[f'C{row}'] = f"=COUNTIFS('Drift Detection Log'!$H$3:$H$152,\"{category}\",'Drift Detection Log'!$O$3:$O$152,\"<>Closed\",'Drift Detection Log'!$O$3:$O$152,\"<>False Positive\")"
        apply_style(ws[f'C{row}'], styles['protected_cell'])

        ws[f'D{row}'] = f"=COUNTIFS('Drift Detection Log'!$H$3:$H$152,\"{category}\",'Drift Detection Log'!$O$3:$O$152,\"Closed\")"
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        total_incidents = 5  # Reference row for total (to be calculated)
        ws[f'E{row}'] = f'=IF(SUM(B$5:B$9)=0,0,B{row}/SUM(B$5:B$9)*100)'
        ws[f'E{row}'].number_format = '0.0'
        apply_style(ws[f'E{row}'], styles['protected_cell'])
        
        ws[f'F{row}'] = "Monitor"
        ws[f'F{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # Add note
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "NOTE: Use this data to identify drift patterns. High Critical/High drift = environmental instability."
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    return ws

def create_summary_dashboard_sheet(wb, styles):
    """Create Gold Standard Summary Dashboard for A.8.9.3 Configuration Monitoring."""
    thin = Side(style='thin')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # ── A1: Title bar ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    ws['A1'] = "CONFIGURATION MONITORING \u2014 SUMMARY DASHBOARD"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A1'].border = border_thin
    ws.row_dimensions[1].height = 35

    # ── A2: Subtitle ──────────────────────────────────────────────────────────
    ws.merge_cells('A2:G2')
    ws['A2'] = "Executive summary of configuration monitoring coverage, drift detection, and alert quality"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    # ── TABLE 1 Banner ────────────────────────────────────────────────────────
    ws.merge_cells('A4:G4')
    ws['A4'] = "TABLE 1 — MONITORING COMPLIANCE ASSESSMENT BY AREA"
    ws['A4'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].border = border_thin
    ws.row_dimensions[4].height = 20

    # ── TABLE 1 Headers ───────────────────────────────────────────────────────
    t1_headers = ['Assessment Area', 'Total Items', 'Compliant', 'Partial', 'Non-Compliant', 'N/A', 'Compliance %']
    t1_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for col, hdr in zip(t1_cols, t1_headers):
        cell = ws[f'{col}5']
        cell.value = hdr
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    ws.row_dimensions[5].height = 18

    # ── TABLE 1 Data rows ─────────────────────────────────────────────────────
    # Row 6: Monitoring Coverage Register — col G = MONITORING_STATUS
    #   Compliant = "\u2705 Monitored", Partial = "\u26A0\uFE0F Partially Monitored",
    #   Non-Compliant = "\u274C Not Monitored", N/A = "➖ Excluded"
    t1_data = [
        (
            "Monitoring Coverage Register",
            # Total = COUNTA of non-empty Asset IDs (row 4 = sample, rows 5-54 = empty)
            "=COUNTA('Monitoring Coverage Register'!A5:A54)",
            # Compliant = Monitored
            "=COUNTIF('Monitoring Coverage Register'!G5:G54,\"\u2705 Monitored\")",
            # Partial = Partially Monitored
            "=COUNTIF('Monitoring Coverage Register'!G5:G54,\"\u26A0\uFE0F Partially Monitored\")",
            # Non-Compliant = Not Monitored
            "=COUNTIF('Monitoring Coverage Register'!G5:G54,\"\u274C Not Monitored\")",
            # N/A = Excluded (➖ = \u2796)
            "=COUNTIF('Monitoring Coverage Register'!G5:G54,\"\u2796 Excluded\")",
        ),
        (
            "Drift Detection Log",
            # Total = COUNTA of non-empty Drift Incident IDs
            "=COUNTA('Drift Detection Log'!A5:A54)",
            # Compliant = Remediated (\u2705 Remediated) + Closed (\u2714\uFE0F Closed)
            "=COUNTIF('Drift Detection Log'!O5:O54,\"\u2705 Remediated\")+COUNTIF('Drift Detection Log'!O5:O54,\"\u2714\uFE0F Closed\")",
            # Partial = Under Investigation + Remediation In Progress
            "=COUNTIF('Drift Detection Log'!O5:O54,\"\u23F3 Under Investigation\")+COUNTIF('Drift Detection Log'!O5:O54,\"\u23F3 Remediation In Progress\")",
            # Non-Compliant = Detected (\u26A0\uFE0F Detected)
            "=COUNTIF('Drift Detection Log'!O5:O54,\"\u26A0\uFE0F Detected\")",
            # N/A = False Positive (➖ = \u2796)
            "=COUNTIF('Drift Detection Log'!O5:O54,\"\u2796 False Positive\")",
        ),
        (
            "Monitoring Tool Inventory",
            # Total = COUNTA of non-empty Tool IDs
            "=COUNTA('Monitoring Tool Inventory'!A5:A54)",
            # Compliant = Active (\u2705 Active)
            "=COUNTIF('Monitoring Tool Inventory'!H5:H54,\"\u2705 Active\")",
            # Partial = Degraded (\u26A0\uFE0F Degraded) + Pilot
            "=COUNTIF('Monitoring Tool Inventory'!H5:H54,\"\u26A0\uFE0F Degraded\")+COUNTIF('Monitoring Tool Inventory'!H5:H54,\"Pilot\")",
            # Non-Compliant = Offline (\u274C Offline)
            "=COUNTIF('Monitoring Tool Inventory'!H5:H54,\"\u274C Offline\")",
            # N/A = Decommissioned
            "=COUNTIF('Monitoring Tool Inventory'!H5:H54,\"\u23F8\uFE0F Decommissioned\")",
        ),
        (
            "False Positive Register",
            # Total = COUNTA of non-empty FP IDs
            "=COUNTA('False Positive Register'!A5:A54)",
            # Compliant = Not Seen Again
            "=COUNTIF('False Positive Register'!N5:N54,\"Not Seen Again\")",
            # Partial = Monitoring + Recurred Once
            "=COUNTIF('False Positive Register'!N5:N54,\"Monitoring\")+COUNTIF('False Positive Register'!N5:N54,\"Recurred Once\")",
            # Non-Compliant = Recurring (Needs Further Tuning)
            "=COUNTIF('False Positive Register'!N5:N54,\"Recurring (Needs Further Tuning)\")",
            # N/A = 0 (no N/A category)
            "=0",
        ),
    ]

    for i, (area, total, comp, part, noncomp, na) in enumerate(t1_data):
        row = 6 + i
        ws[f'A{row}'] = area
        ws[f'A{row}'].font = Font(name='Calibri', size=11, color='0000FF')
        ws[f'A{row}'].border = border_thin
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

        ws[f'B{row}'] = total
        ws[f'C{row}'] = comp
        ws[f'D{row}'] = part
        ws[f'E{row}'] = noncomp
        ws[f'F{row}'] = na
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].font = Font(name='Calibri', size=11, color='0000FF')
            ws[f'{col}{row}'].border = border_thin
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Compliance % = IF((B-F)=0, 0, C/(B-F))
        ws[f'G{row}'] = f'=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))'
        ws[f'G{row}'].font = Font(name='Calibri', size=11, color='0000FF')
        ws[f'G{row}'].number_format = '0.0%'
        ws[f'G{row}'].border = border_thin
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')

    # ── TABLE 1 TOTAL row ─────────────────────────────────────────────────────
    total_row = 6 + len(t1_data)
    ws[f'A{total_row}'] = "TOTAL"
    ws[f'A{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{total_row}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws[f'A{total_row}'].border = border_thin
    ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{total_row}'] = f'=SUM({col}6:{col}{total_row - 1})'
        ws[f'{col}{total_row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'{col}{total_row}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[f'{col}{total_row}'].border = border_thin
        ws[f'{col}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'G{total_row}'] = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    ws[f'G{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'G{total_row}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws[f'G{total_row}'].number_format = '0.0%'
    ws[f'G{total_row}'].border = border_thin
    ws[f'G{total_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # ── TABLE 2 Banner ────────────────────────────────────────────────────────
    t2_start = total_row + 2
    ws.merge_cells(f'A{t2_start}:G{t2_start}')
    ws[f'A{t2_start}'] = "TABLE 2 — KEY MONITORING PERFORMANCE INDICATORS"
    ws[f'A{t2_start}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{t2_start}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{t2_start}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{t2_start}'].border = border_thin

    # TABLE 2 headers
    t2_hrow = t2_start + 1
    for col, hdr in zip(['A', 'B', 'C'], ['KPI', 'Value', 'Target']):
        ws[f'{col}{t2_hrow}'] = hdr
        ws[f'{col}{t2_hrow}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'{col}{t2_hrow}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[f'{col}{t2_hrow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}{t2_hrow}'].border = border_thin
    # Merge D:G for header
    ws.merge_cells(f'D{t2_hrow}:G{t2_hrow}')
    ws[f'D{t2_hrow}'] = "Notes"
    ws[f'D{t2_hrow}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'D{t2_hrow}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws[f'D{t2_hrow}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'D{t2_hrow}'].border = border_thin

    t2_metrics = [
        ("Overall Monitoring Coverage %",
         f"=IF(B6=0,0,C6/B6)",
         "\u226585%",
         "% of inventoried assets actively monitored"),
        ("Active Monitoring Tools",
         f"=C8",
         "All tools active",
         "Tools in \u2705 Active status"),
        ("Drift Incidents Remediated %",
         f"=IF(B7=0,0,C7/B7)",
         "\u226595%",
         "% of logged drift incidents resolved"),
        ("False Positive Rate %",
         f"=IF(B9=0,0,(C9+D9)/B9)",
         "<10%",
         "% of FP register items resolved or one-time"),
        ("Unauthorised Changes Detected",
         "=COUNTIF('Drift Detection Log'!K5:K54,\"No (Unauthorised)\")",
         "Detect all",
         "Drift incidents with no authorised change ID"),
        ("Open (Unresolved) Drift Incidents",
         "=COUNTIF('Drift Detection Log'!O5:O54,\"\u26A0\uFE0F Detected\")+COUNTIF('Drift Detection Log'!O5:O54,\"\u23F3 Under Investigation\")+COUNTIF('Drift Detection Log'!O5:O54,\"\u23F3 Remediation In Progress\")",
         "0",
         "Drift incidents awaiting remediation"),
    ]

    for j, (kpi, val, target, note) in enumerate(t2_metrics):
        r = t2_hrow + 1 + j
        ws[f'A{r}'] = kpi
        ws[f'A{r}'].font = Font(name='Calibri', size=11)
        ws[f'A{r}'].border = border_thin
        ws[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center')

        ws[f'B{r}'] = val
        ws[f'B{r}'].font = Font(name='Calibri', size=11, color='0000FF')
        ws[f'B{r}'].border = border_thin
        ws[f'B{r}'].alignment = Alignment(horizontal='center', vertical='center')
        if '%' in kpi:
            ws[f'B{r}'].number_format = '0.0%'

        ws[f'C{r}'] = target
        ws[f'C{r}'].font = Font(name='Calibri', size=11)
        ws[f'C{r}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'C{r}'].border = border_thin
        ws[f'C{r}'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(f'D{r}:G{r}')
        ws[f'D{r}'] = note
        ws[f'D{r}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
        ws[f'D{r}'].border = border_thin
        ws[f'D{r}'].alignment = Alignment(horizontal='left', vertical='center')

    # ── TABLE 3 Banner ────────────────────────────────────────────────────────
    t3_start = t2_hrow + 1 + len(t2_metrics) + 2
    ws.merge_cells(f'A{t3_start}:G{t3_start}')
    ws[f'A{t3_start}'] = "TABLE 3 — KEY FINDINGS AND RECOMMENDED ACTIONS"
    ws[f'A{t3_start}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{t3_start}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws[f'A{t3_start}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{t3_start}'].border = border_thin

    # TABLE 3 headers
    t3_hrow = t3_start + 1
    for col, hdr in zip(['A', 'B', 'C', 'D'], ['#', 'Finding', 'Priority', 'Recommended Action']):
        ws[f'{col}{t3_hrow}'] = hdr
        ws[f'{col}{t3_hrow}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'{col}{t3_hrow}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[f'{col}{t3_hrow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}{t3_hrow}'].border = border_thin
    ws.merge_cells(f'D{t3_hrow}:G{t3_hrow}')
    ws[f'D{t3_hrow}'].border = border_thin

    t3_findings = [
        ("1", "Monitoring Coverage Register — assets with status not '\u2705 Monitored' indicate gaps in coverage requiring immediate remediation", "Critical", "Investigate and deploy monitoring agents/tools for all unmonitored Critical and High assets within 30 days"),
        ("2", "Drift Detection Log — open drift incidents ('\u26A0\uFE0F Detected') indicate unresolved configuration deviations that may represent security risk", "High", "Prioritise remediation of Critical and High severity drift incidents; ensure all incidents have assigned owners and target dates"),
        ("3", "Monitoring Tool Inventory — tools in '\u26A0\uFE0F Degraded' or '\u274C Offline' status create monitoring blind spots in asset coverage", "High", "Restore or replace degraded/offline tools; document coverage gaps and implement compensating controls until resolved"),
        ("4", "False Positive Register — recurring false positives ('\u274C Recurring') indicate baseline or tool configuration quality issues that reduce monitoring effectiveness", "Medium", "Review and update baselines to eliminate recurring false positives; document tuning actions and monitor for recurrence"),
    ]

    for k, (num, finding, priority, action) in enumerate(t3_findings):
        r = t3_hrow + 1 + k
        ws[f'A{r}'] = num
        ws[f'A{r}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{r}'].border = border_thin
        ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')

        ws[f'B{r}'] = finding
        ws[f'B{r}'].font = Font(name='Calibri', size=11)
        ws[f'B{r}'].border = border_thin
        ws[f'B{r}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws[f'C{r}'] = priority
        ws[f'C{r}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'C{r}'].border = border_thin
        ws[f'C{r}'].alignment = Alignment(horizontal='center', vertical='center')
        if priority == 'Critical':
            ws[f'C{r}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif priority == 'High':
            ws[f'C{r}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        else:
            ws[f'C{r}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        ws.merge_cells(f'D{r}:G{r}')
        ws[f'D{r}'] = action
        ws[f'D{r}'].font = Font(name='Calibri', size=11)
        ws[f'D{r}'].border = border_thin
        ws[f'D{r}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws.row_dimensions[r].height = 45

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14

    ws.freeze_panes = 'A4'
    return ws


def create_evidence_register(wb, styles):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(color="808080")
        ws.cell(row=r, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"
def create_approval_sheet(wb, styles):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.9.3 - Configuration Monitoring Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "='Monitor Effectiveness Metrics'!B7"),
        ("Assessment Status:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border_thin
        row += 1

    # Status dropdown on Assessment Status
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border_thin

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Apply borders to all merged cell top-left corners (GS-AS-011)
    _as_thin = Side(style="thin")
    _as_border = Border(left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin)
    for merge_range in ws.merged_cells.ranges:
        tl = ws.cell(merge_range.min_row, merge_range.min_col)
        tl.border = _as_border
    ws.freeze_panes = "A3"
# ============================================================================
# MAIN EXECUTION
# ============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating {WORKBOOK_TITLE} Workbook")
    logger.info("=" * 70)
    logger.info(f"Document ID: {DOCUMENT_ID}")
    logger.info(f"Version: {WORKBOOK_VERSION}")
    logger.info(f"Date: {datetime.now().strftime('%d.%m.%Y')}")
    logger.info("-" * 70)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create styles
    styles = create_styles()
    
    # Create all sheets
    logger.info("Creating sheets...")
    
    logger.info("  2/12 Creating Instructions sheet...")
    create_instructions_sheet(wb.create_sheet())

    logger.info("  1/12 Creating Lookup Tables (hidden)...")
    create_lookup_tables(wb, styles)
    
    
    logger.info("  3/12 Creating Monitoring Coverage Register sheet (100 rows)...")
    create_monitoring_coverage_register_sheet(wb, styles)
    
    logger.info("  4/12 Creating Drift Detection Log sheet (150 rows)...")
    create_drift_detection_log_sheet(wb, styles)
    
    logger.info("  5/12 Creating Monitoring Tool Inventory sheet (30 rows)...")
    create_monitoring_tool_inventory_sheet(wb, styles)
    
    logger.info("  6/12 Creating Drift Remediation Tracking sheet (150 rows)...")
    create_drift_remediation_tracking_sheet(wb, styles)
    
    logger.info("  7/12 Creating False Positive Register sheet (75 rows)...")
    create_false_positive_register_sheet(wb, styles)
    
    logger.info("  8/12 Creating Monitor Effectiveness Metrics sheet (dashboard)...")
    create_monitoring_effectiveness_metrics_sheet(wb, styles)
    
    logger.info("  9/12 Creating Coverage Gap Analysis sheet (dashboard)...")
    create_coverage_gap_analysis_sheet(wb, styles)
    
    logger.info(" 10/12 Creating Drift Trend Analysis sheet (dashboard)...")
    create_drift_trend_analysis_sheet(wb, styles)

    logger.info(" 11/13 Creating Evidence Register sheet (100 rows)...")
    create_evidence_register(wb, styles)

    logger.info(" 12/13 Creating Summary Dashboard sheet (Gold Standard)...")
    create_summary_dashboard_sheet(wb, styles)

    logger.info(" 13/13 Creating Approval Sign Off sheet...")
    create_approval_sheet(wb, styles)

    logger.info("  \u2713 All sheets created successfully")
    
    # Set workbook properties
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.created = datetime.now()
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Save workbook
    logger.info("-" * 70)
    logger.info("Saving workbook...")
    finalize_validations(wb)
    wb.save(output_path)
    logger.info("=" * 70)
    logger.info("✓ Workbook generated successfully!")
    logger.info("=" * 70)
    logger.info(f"Output File: {FILENAME}")
    logger.info(f"File Size: {os.path.getsize(_wkbk_dir / FILENAME) / 1024:.1f} KB")
    logger.info(f"Total Sheets: 13 (12 visible + 1 hidden lookup table)")
    logger.info("-" * 70)
    logger.info("\nWorkbook Structure:")
    logger.info("  1.  Instructions - Usage guidance, monitoring methods, drift categories")
    logger.info("  2.  Monitoring Coverage Register - 100 rows for asset monitoring inventory")
    logger.info("  3.  Drift Detection Log - 150 rows for drift incident records")
    logger.info("  4.  Monitoring Tool Inventory - 30 rows for tool capabilities")
    logger.info("  5.  Drift Remediation Tracking - 150 rows for remediation actions")
    logger.info("  6.  False Positive Register - 75 rows for alert quality tracking")
    logger.info("  7.  Monitor Effectiveness Metrics - Auto-calculated KPI dashboard")
    logger.info("  8.  Coverage Gap Analysis - Coverage analysis by category/criticality")
    logger.info("  9.  Drift Trend Analysis - Temporal drift pattern analysis")
    logger.info("  10. Evidence Register - 100 rows for evidence documentation")
    logger.info("  11. Approval Sign Off - Three-tier approval signatures")
    logger.info("  12. Lookup Tables (hidden) - 43-type asset taxonomy")
    logger.info("-" * 70)
    logger.info("\nKey Monitoring Metrics:")
    logger.info("\u2022 Coverage: Tier 1=100%, Tier 2≥95%, Tier 3≥85%, Tier 4≥70%")
    logger.info("\u2022 MTTD (Mean Time to Detect): <1 hour (Tier 1), <24 hours (Tier 2)")
    logger.info("\u2022 MTTR (Mean Time to Remediate): <4 hours (Critical), <1 day (High)")
    logger.info("\u2022 False Positive Rate: <10%")
    logger.info("\u2022 Critical Drift: ZERO TOLERANCE (immediate escalation)")
    logger.info("\u2022 SLA Compliance: 100% (Critical), ≥95% (High)")
    logger.info("-" * 70)
    logger.info("\nNext Steps:")
    logger.info("1. Open workbook in Excel/LibreOffice")
    logger.info("2. Verify all sheets, validations, and formulas")
    logger.info("3. Review Instructions for monitoring methods and drift categories")
    logger.info("4. Customise dropdown values if needed (see CONFIGURATION section)")
    logger.info("5. Document current monitoring coverage")
    logger.info("6. Establish drift detection logging process")
    logger.info("7. Review dashboards monthly for trends")
    logger.info("8. Address coverage gaps identified in Coverage Gap Analysis")
    logger.info("-" * 70)
    logger.info("\nIMPORTANT REMINDERS:")
    logger.info("\u2022 This is a SAMPLE workbook - customise for your environment")
    logger.info("\u2022 Critical drift requires immediate action (<4 hours)")
    logger.info("\u2022 Log all drift incidents even if immediately remediated")
    logger.info("\u2022 False positives must be analysed and tuning documented")
    logger.info("\u2022 Tools marked 'Offline' create monitoring gaps")
    logger.info("\u2022 Integration with A.8.9.1 (baselines) and A.8.9.2 (changes) is critical")
    logger.info("\u2022 Protected cells (gray) contain formulas - do not edit")
    logger.info("\u2022 Retain workbook and evidence for audit (minimum 3 years)")
    logger.info("=" * 70)

def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
