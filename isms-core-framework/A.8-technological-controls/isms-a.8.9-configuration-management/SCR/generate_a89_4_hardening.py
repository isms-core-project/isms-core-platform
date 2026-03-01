#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.9.4 - Security Hardening Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Assessment Domain 4 of 4: Security Hardening and Compliance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific hardening standards, compliance scanning tools,
and gap remediation processes.

Key customisation areas:
1. Hardening standards selection (CIS, STIG, vendor guides per your stack)
2. Compliance scanning tools (match your actual deployment)
3. Gap remediation priorities (adapt to your risk appetite)
4. Exception approval authorities (align with organisational roles)
5. Compliance thresholds (based on your security requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMISE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
security hardening implementation and compliance verification against ISO
27001:2022 Control A.8.9 requirements.

**Purpose:**
Enables systematic assessment of hardening standard selection, implementation,
compliance verification, gap analysis, and remediation tracking to ensure
secure system configurations across all asset types.

**Assessment Scope:**
- Hardening standards selection and applicability mapping
- Compliance scanning tool deployment and coverage
- Gap analysis and risk assessment
- Hardening exception management and approval
- Implementation evidence and validation
- Compliance reporting and trending
- Remediation tracking and SLA compliance
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and hardening standards
2. Hardening Standards - Standard selection and asset type mapping
3. Compliance Scans - Automated scan results and compliance rates
4. Gap Analysis - Identified gaps, risk assessment, remediation planning
5. Hardening Exceptions - Approved exceptions and compensating controls
6. Implementation Evidence - Pre/post hardening validation
7. Compliance Reports - Quarterly compliance trending and analysis
8. Gap Analysis Summary - Consolidated gaps and remediation requirements
9. Evidence Register - Audit evidence tracking (100+ entries)
10. Approval & Sign-Off - Three-tier approval workflow

**Key Features:**
- Data validation with hardening standard and asset type dropdowns
- Conditional formatting for compliance status and gap severity
- Automated gap identification from scan results
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with compliance scanning tools

**Integration:**
consolidates data from all four configuration management assessment domains
for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_4_hardening.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_4_hardening.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_4_hardening.py --date 20250127

Output:
    File: ISMS-IMP-A.8.9.4_Security_Hardening_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise hardening standards by asset type
    2. Document compliance scanning tool deployment
    3. Import compliance scan results
    4. Conduct gap analysis with risk assessment
    5. Document hardening exceptions with compensating controls
    6. Collect pre/post hardening validation evidence
    7. Generate quarterly compliance trending reports
    8. Define remediation actions with timelines and owners
    9. Collect and link audit evidence (scan reports, hardening scripts)
    10. Obtain three-tier stakeholder approvals

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Assessment Domain:    4 of 4 (Security Hardening and Compliance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-POL-A.8.9, Section 2.5: Security Hardening & Compliance
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9, Part 1: Technical Standards Reference
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.1: Baseline Configuration Assessment (Domain 1)
    - ISMS-IMP-A.8.9.2: Change Control Assessment (Domain 2)
    - ISMS-IMP-A.8.9.3: Configuration Monitoring Assessment (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.9.4 specification
    - Supports comprehensive security hardening evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Hardening Standards:**
Security hardening benchmarks evolve continuously. Review CIS Benchmarks
quarterly for updates, monitor DISA STIG releases, and track vendor security
guide revisions. Deprecated hardening controls must be identified and updated.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of hardening implementation, compliance
scanning, and gap remediation.

**Data Protection:**
Assessment workbooks contain sensitive security details including:
- Hardening compliance scan results with gap details
- Security vulnerability information
- Exception justifications and compensating controls
- System configuration weaknesses

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check hardening compliance rates and gap remediation progress
- Semi-annually: Update hardening standards for benchmark revisions
- Annually: Complete reassessment of all asset types
- Ad-hoc: When new vulnerabilities or hardening guidance published

**Quality Assurance:**
Have security architects, hardening specialists, and compliance officers
validate assessments before using results for reporting or remediation
prioritization.

**Regulatory Alignment:**
Ensure hardening practices align with applicable requirements:
- ISO 27001:2022: Control A.8.9 secure configuration
- CIS Controls v8.1: Control 4 secure configuration requirements
- PCI DSS v4.0.1: System hardening and configuration standards
- Sector-specific: Regulatory hardening requirements
- Internal: Organisational security hardening policies

Customize assessment criteria to include regulatory-specific requirements.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import os
import re
import sys
from datetime import datetime, timedelta
from typing import List, Dict, Tuple

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


# =============================================================================
# =============================================================================

# CUSTOMISE: Assessment metadata

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Document identification

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.9.4"
WORKBOOK_NAME    = "Security Hardening Assessment"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
CONTROL_ID   = "A.8.9"
CONTROL_NAME = "Configuration Management"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
WORKBOOK_VERSION = "1.0"

# CUSTOMISE: Compliance targets for different asset tiers
COMPLIANCE_TARGETS = {
    'Critical': 100,   # Critical assets: 100% compliance required
    'High': 98,        # High assets: 98% compliance target
    'Medium': 95,      # Medium assets: 95% compliance target
    'Low': 90,         # Low assets: 90% compliance target
    'Overall': 95,     # Overall organisation target: 95%
}

# CUSTOMISE: Risk scoring weights (adjust based on your risk model)
RISK_WEIGHTS = {
    'asset_tier': {
        'Critical': 5,
        'High': 4,
        'Medium': 3,
        'Low': 2,
    },
    'control_severity': {
        'Critical': 5,
        'High': 4,
        'Medium': 3,
        'Low': 2,
    },
    'exploitation_likelihood': {
        'Very High': 5,
        'High': 4,
        'Medium': 3,
        'Low': 2,
        'Very Low': 1,
    },
}

# CUSTOMISE: Remediation SLA targets (days) based on priority
REMEDIATION_SLA = {
    'Critical': 3,      # Critical gaps: 3 days
    'High': 7,          # High gaps: 7 days
    'Medium': 30,       # Medium gaps: 30 days
    'Low': 90,          # Low gaps: 90 days
}

# 43-type asset taxonomy (consistent with A.8.9.1, A.8.9.2, A.8.9.3)
# DO NOT MODIFY - This taxonomy must remain consistent across all A.8.9 assessments
ASSET_TAXONOMY = {
    'Infrastructure': [
        'Physical Server',
        'Virtual Server (VM)',
        'Hypervisor/VM Host',
        'Container Host',
        'Storage Array',
        'Backup Appliance',
        'Database Server',
        'Application Server',
        'Legacy System',
    ],
    'Endpoint': [
        'Corporate Workstation',
        'Corporate Laptop',
        'Executive Device',
        'Mobile Device (Corporate)',
        'Mobile Device (BYOD)',
        'Thin Client',
        'Kiosk/Shared Terminal',
    ],
    'Network Services': [
        'Core Router',
        'Edge Router',
        'L2/L3 Switch',
        'Wireless Access Point',
        'Firewall',
        'IDS/IPS Appliance',
        'Load Balancer',
        'VPN Concentrator',
    ],
    'Applications': [
        'Web Application',
        'API Service',
        'Database Management System',
        'File Server/NAS',
        'Email System',
        'Authentication Service (IAM)',
        'Monitoring/Logging System',
    ],
    'Cloud & Virtual': [
        'IaaS Instance',
        'PaaS Application',
        'SaaS Application Configuration',
        'Container (Docker/K8s)',
        'Serverless Function',
        'Cloud Storage Bucket',
        'Cloud Network Configuration',
    ],
    'IoT & OT': [
        'Building Management System',
        'Physical Security System',
        'Industrial Control System (ICS)',
        'SCADA System',
        'IoT Sensor/Device',
    ],
}

# Flatten asset taxonomy for dropdowns
ALL_ASSET_TYPES = []
for category, types in ASSET_TAXONOMY.items():
    ALL_ASSET_TYPES.extend(types)

# Dropdown value definitions
DROPDOWNS = {
    'asset_tier': ['Critical', 'High', 'Medium', 'Low'],
    'standard_category': [
        'Industry Benchmark',
        'Government Standard',
        'Regulatory Requirement',
        'Vendor Baseline',
        'Framework Control',
        'Custom Organisational Standard',
    ],
    'compliance_level': ['Level 1', 'Level 2', 'Custom'],
    'mandatory_optional': ['Mandatory', 'Optional'],
    'review_frequency': ['Monthly', 'Quarterly', 'Semi-Annual', 'Annual'],
    'status': ['\u2705 Active', '\u26A0\uFE0F Deprecated', 'Planned'],
    'applicability': ['Required', 'Recommended', 'Optional', 'Not Applicable'],
    'compliance_status_asset': [
        '\u2705 Fully Compliant',
        '\u2705 Compliant',
        '\u26A0\uFE0F Substantially Compliant',
        '\u26A0\uFE0F Partially Compliant',
        '\u274C Non-Compliant',
    ],
    'remediation_status': [
        '❌ Not Started',
        'Planning',
        '⏳ In Progress',
        '\u274C Blocked',
        '\u2705 Completed',
        '\u26A0\uFE0F Accepted as Exception',
    ],
    'implementation_status': [
        '\u2705 Implemented',
        '\u26A0\uFE0F Partial',
        '\u274C Not Implemented',
        '➖ Not Applicable',
        'Planned',
        '⏳ In Progress',
    ],
    'implementation_method': [
        'Automated Tool',
        'Manual Configuration',
        'Native Feature',
        'Third-Party Tool',
        'Compensating Control',
        'Not Implemented',
    ],
    'control_category': [
        'Access Control',
        'Audit & Logging',
        'Authentication',
        'Network Security',
        'Data Protection',
        'System Hardening',
        'Patch Management',
        'Secure Configuration',
        'Service Minimization',
        'Physical Security',
        'Cryptography',
        'Backup & Recovery',
    ],
    'control_severity': ['Critical', 'High', 'Medium', 'Low'],
    'compliance_status_control': ['\u2705 Pass', '\u274C Fail', '\u26A0\uFE0F Partial', '➖ N/A'],
    'verification_method': [
        'Configuration Export',
        'Security Tool Report',
        'Manual Inspection',
        'Audit Log Review',
        'Test/Validation',
        'Documentation Review',
    ],
    'yes_no': ['Yes', 'No'],
    'exception_type': [
        'Technical Limitation',
        'Business Requirement',
        'Legacy System',
        'Performance Impact',
        'Vendor Limitation',
        'Cost Prohibitive',
        'Temporary Transition',
        'Regulatory Conflict',
    ],
    'exception_status': [
        '⏳ Pending Review',
        '⏳ Under Review',
        '\u2705 Approved',
        '\u26A0\uFE0F Conditionally Approved',
        '\u274C Rejected',
        '⏰ Expired',
        '\u2705 Closed',
    ],
    'exception_duration': ['3 Months', '6 Months', '12 Months', '24 Months', 'Indefinite'],
    'residual_risk': ['Critical', 'High', 'Medium', 'Low'],
    'compensating_effectiveness': ['Full', 'Partial', 'None'],
    'gap_type': [
        'Hardening Gap',
        'Configuration Drift',
        'Vulnerability',
        'Audit Finding',
        'Threat Response',
    ],
    'discovery_method': [
        'Security Assessment',
        'Vulnerability Scan',
        'Configuration Monitoring',
        'Penetration Test',
        'Incident Response',
        'Audit',
        'Threat Intelligence',
    ],
    'remediation_strategy': [
        'Configuration Change',
        'Software Update',
        'Policy Change',
        'Compensating Control',
        'Accept as Exception',
        'Asset Decommission',
        'Defer',
    ],
    'estimated_effort': [
        '<1 hour',
        '1-4 hours',
        '1 day',
        '2-5 days',
        '1-2 weeks',
        '2-4 weeks',
        '>1 month',
    ],
    'remediation_priority': ['Critical', 'High', 'Medium', 'Low'],
    'remediation_tracking_status': [
        'Identified',
        'Planning',
        '\u2705 Approved',
        '⏳ In Progress',
        '⏳ Verification',
        '\u2705 Completed',
        '\u274C Blocked',
        '\u26A0\uFE0F Deferred',
        '\u2705 Closed - Fixed',
        '\u26A0\uFE0F Closed - Exception',
        '➖ Closed - Not Required',
    ],
    'verification_result': ['\u2705 Pass', '\u274C Fail'],
    'evidence_type': [
        'Configuration Export',
        'Screenshot',
        'Security Scan Report',
        'Audit Log Extract',
        'Policy/Procedure Document',
        'Test Results',
        'Attestation',
        'Change Record',
        'Exception Approval',
        'Vendor Documentation',
    ],
    'evidence_source': [
        'Automated Tool',
        'Manual Collection',
        'System Export',
        'Security Tool',
        'Change Management System',
        'Documentation Repository',
        'Email',
        'Meeting Minutes',
    ],
    'collection_method': [
        'API/Script',
        'Command Line',
        'GUI Screenshot',
        'File Export',
        'Report Generation',
        'Manual Documentation',
        'Copy/Paste',
    ],
    'evidence_validity': [
        'Point-in-Time',
        '1 Month',
        '3 Months',
        '6 Months',
        '12 Months',
        'Continuous',
        'Until Changed',
    ],
    'evidence_status': ['\u2705 Active', '⏰ Expired', '\u26A0\uFE0F Superseded'],
    'exploitation_likelihood': ['Very High', 'High', 'Medium', 'Low', 'Very Low'],
}

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

# Fonts
FONT_HEADER = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
FONT_SUBHEADER = Font(name='Calibri', size=12, bold=True)
FONT_NORMAL = Font(name='Calibri', size=11)
FONT_SMALL = Font(name='Calibri', size=10)
FONT_BOLD = Font(name='Calibri', size=11, bold=True)

# Fills
FILL_HEADER = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
FILL_SUBHEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
FILL_LIGHT_BLUE = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
FILL_LIGHT_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_LIGHT_YELLOW = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
FILL_LIGHT_RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_RED = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
FILL_ORANGE = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Standard amber
FILL_GRAY = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FILL_DARK_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Same as FILL_GREEN (standard palette)
FILL_INPUT = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

THICK_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def set_column_widths(ws, widths: Dict[str, float]):
    """
    Set column widths for a worksheet.
    
    Args:
        ws: Worksheet object
        widths: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_header_row(ws, row: int, headers: List[str], start_col: int = 1):
    """
    Apply header formatting to a row.
    
    Args:
        ws: Worksheet object
        row: Row number
        headers: List of header text
        start_col: Starting column (default 1 = A)
    """
    for idx, header in enumerate(headers, start=start_col):
        col_letter = get_column_letter(idx)
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, validation_list: List[str],
                        allow_blank: bool = True, error_title: str = "Invalid Entry",
                        error_message: str = "Please select from the dropdown list.",
                        validations: List = None):
    """
    Add dropdown data validation to a cell range.

    Args:
        ws: Worksheet object
        cell_range: Cell range (e.g., "B2:B100")
        validation_list: List of valid values
        allow_blank: Allow blank cells
        error_title: Error dialog title
        error_message: Error dialog message
        validations: Optional list to collect DV objects for batch application
    """
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(validation_list)}"',
        allow_blank=allow_blank,
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_message
    )
    if validations is not None:
        validations.append(dv)
    else:
        ws.add_data_validation(dv)
    dv.add(cell_range)


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
def add_conditional_formatting_status(ws, cell_range: str, status_col: str):
    """
    Add conditional formatting for status columns.

    Args:
        ws: Worksheet object
        cell_range: Range to apply formatting (e.g. 'W3:W52')
        status_col: Column letter of status field
    """
    # Extract the first row number from the cell_range (e.g. 'W3:W52' -> 3)
    first_row = re.search(r'(\d+)', cell_range).group(1)

    # Green for positive statuses
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("Fully Compliant",${status_col}{first_row})),ISNUMBER(SEARCH("Compliant",${status_col}{first_row})),'
                    f'ISNUMBER(SEARCH("Completed",${status_col}{first_row})),ISNUMBER(SEARCH("Pass",${status_col}{first_row})),ISNUMBER(SEARCH("Approved",${status_col}{first_row})))'],
            fill=FILL_GREEN
        )
    )

    # Yellow for warning statuses
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("Substantially Compliant",${status_col}{first_row})),ISNUMBER(SEARCH("Partial",${status_col}{first_row})),'
                    f'ISNUMBER(SEARCH("In Progress",${status_col}{first_row})),ISNUMBER(SEARCH("Under Review",${status_col}{first_row})))'],
            fill=FILL_YELLOW
        )
    )

    # Red for critical statuses
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("Non-Compliant",${status_col}{first_row})),ISNUMBER(SEARCH("Fail",${status_col}{first_row})),'
                    f'ISNUMBER(SEARCH("Blocked",${status_col}{first_row})),ISNUMBER(SEARCH("Rejected",${status_col}{first_row})),ISNUMBER(SEARCH("Expired",${status_col}{first_row})))'],
            fill=FILL_RED
        )
    )


def generate_auto_id(prefix: str, row_num: int, padding: int = 3) -> str:
    """
    Generate auto-incrementing ID with prefix.
    
    Args:
        prefix: ID prefix (e.g., "HS", "HC", "EXC")
        row_num: Row number (1-based)
        padding: Number of digits for padding
        
    Returns:
        Formatted ID string
    """
    return f"{prefix}-{str(row_num).zfill(padding)}"


def calculate_date_offset(base_date_cell: str, offset_days: int, ws_name: str = None) -> str:
    """
    Generate Excel formula for date calculation.
    
    Args:
        base_date_cell: Cell reference for base date
        offset_days: Days to add
        ws_name: Worksheet name if cross-sheet reference
        
    Returns:
        Excel formula string
    """
    if ws_name:
        return f"={ws_name}!{base_date_cell}+{offset_days}"
    return f"={base_date_cell}+{offset_days}"


# =============================================================================
# SHEET GENERATION FUNCTIONS
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab for applicable systems/services.', '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Check all applicable items in the Compliance Checklist for each section.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_hardening_standard_register(wb: Workbook) -> None:
    """
    Create the Hardening Standard Register sheet.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Hardening Standard Register")
    ws.sheet_view.showGridLines = False
    
    # Define headers
    headers = [
        'Standard ID',
        'Standard Name',
        'Standard Category',
        'Standard Version',
        'Issuing Authority',
        'Applicability Scope',
        'Compliance Level',
        'Mandatory Optional',
        'Regulatory Driver',
        'Control Count',
        'Implementation Target',
        'Review Frequency',
        'Last Review Date',
        'Next Review Date',
        'Standard Owner',
        'Documentation Location',
        'Notes',
        'Status',
    ]
    
    # Add title row
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = "HARDENING STANDARD REGISTER"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Apply headers at row 2
    apply_header_row(ws, 2, headers)

    # Freeze panes below header
    ws.freeze_panes = 'A3'

    # Set column widths
    widths = {
        'A': 12,  # Standard_ID
        'B': 35,  # Standard_Name
        'C': 25,  # Standard_Category
        'D': 15,  # Standard_Version
        'E': 25,  # Issuing_Authority
        'F': 30,  # Applicability_Scope
        'G': 15,  # Compliance_Level
        'H': 18,  # Mandatory_Optional
        'I': 25,  # Regulatory_Driver
        'J': 14,  # Control_Count
        'K': 18,  # Implementation_Target
        'L': 18,  # Review_Frequency
        'M': 18,  # Last_Review_Date
        'N': 18,  # Next_Review_Date
        'O': 25,  # Standard_Owner
        'P': 30,  # Documentation_Location
        'Q': 40,  # Notes
        'R': 12,  # Status
    }
    set_column_widths(ws, widths)

    # Add data validation
    validations = []
    add_data_validation(ws, 'C3:C32', DROPDOWNS['standard_category'], validations=validations)
    add_data_validation(ws, 'G3:G32', DROPDOWNS['compliance_level'], validations=validations)
    add_data_validation(ws, 'H3:H32', DROPDOWNS['mandatory_optional'], validations=validations)
    add_data_validation(ws, 'L3:L32', DROPDOWNS['review_frequency'], validations=validations)
    add_data_validation(ws, 'R3:R32', DROPDOWNS['status'], validations=validations)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Add formulas for 30 rows
    for row in range(3, 33):
        # Standard_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "HS-"&TEXT(ROW()-2,"000"), "")'
        ws[f'A{row}'].font = FONT_NORMAL

        # Implementation_Target default 95%
        ws[f'K{row}'].value = 0.95
        ws[f'K{row}'].number_format = '0%'

        # Review_Frequency default Quarterly
        ws[f'L{row}'].value = 'Quarterly'

        # Status default Active
        ws[f'R{row}'].value = 'Active'

        # Mandatory_Optional default Mandatory
        ws[f'H{row}'].value = 'Mandatory'

    # Conditional formatting
    # Next_Review_Date past due: Red
    ws.conditional_formatting.add(
        'N3:N32',
        CellIsRule(operator='lessThan', formula=['TODAY()'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )

    # Next_Review_Date within 30 days: Yellow
    ws.conditional_formatting.add(
        'N3:N32',
        CellIsRule(operator='between', formula=['TODAY()', 'TODAY()+30'], fill=FILL_YELLOW)
    )

    # Status = Deprecated: Gray
    ws.conditional_formatting.add(
        'A3:R32',
        FormulaRule(formula=['ISNUMBER(SEARCH("Deprecated",$R3))'], fill=FILL_GRAY)
    )

    # Mandatory + Target <95%: Orange
    ws.conditional_formatting.add(
        'K3:K32',
        FormulaRule(formula=['AND($H3="Mandatory", $K3<0.95)'], fill=FILL_ORANGE)
    )

    # Unlock data entry cells

    # Protect sheet

def create_asset_type_hardening_matrix(wb: Workbook) -> None:
    """
    Create the Asset Type Hardening Matrix sheet.
    
    This matrix maps hardening standards (columns) to asset types (rows),
    indicating which standards apply to which asset categories.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Asset Type Hardening Matrix")
    ws.sheet_view.showGridLines = False

    # CUSTOMISE: Number of standards you expect
    # This creates space for 30 standards - adjust if needed
    NUM_STANDARDS = 30

    # Summary columns after standards
    summary_col_start = NUM_STANDARDS + 2
    summary_headers = [
        'Required Standards Count',
        'Recommended Standards Count',
        'Total Applicable Standards',
        'High Hardening Burden',
    ]

    # Row 1: Merged title row
    last_col = get_column_letter(summary_col_start + len(summary_headers) - 1)
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = "ASSET TYPE HARDENING MATRIX"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Row 2: Column headers
    ws['A2'] = 'ASSET TYPE'
    ws['A2'].font = FONT_HEADER
    ws['A2'].fill = FILL_HEADER
    ws['A2'].alignment = ALIGN_CENTER
    ws['A2'].border = THIN_BORDER

    # Columns B through AE (30 columns): Standard headers
    # These will pull Standard_ID from Hardening Standard Register
    for col_idx in range(2, NUM_STANDARDS + 2):  # B to AE
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}2']

        # Formula to pull Standard_ID from register
        # Shows "HS-001: Standard Name" format
        # HSR data starts at row 3 now, so standard N is at row col_idx+1
        cell.value = f"=IF('Hardening Standard Register'!A{col_idx+1}<>\"\", " \
                    f"'Hardening Standard Register'!A{col_idx+1}&\": \"&" \
                    f"LEFT('Hardening Standard Register'!B{col_idx+1},20), \"\")"
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = 18

    for idx, header in enumerate(summary_headers):
        col_letter = get_column_letter(summary_col_start + idx)
        cell = ws[f'{col_letter}2']
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_SUBHEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = 22

    # Set Asset Type column width
    ws.column_dimensions['A'].width = 30

    # Freeze panes below header
    ws.freeze_panes = 'A3'

    # Row 3 onwards: Asset types (43 rows)
    row = 3
    for category, asset_types in ASSET_TAXONOMY.items():
        # Category header row (merged across all columns)
        ws[f'A{row}'] = f'--- {category.upper()} ---'
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, italic=True)
        ws[f'A{row}'].fill = FILL_LIGHT_BLUE
        ws[f'A{row}'].alignment = ALIGN_LEFT
        
        # Merge category header across all standard columns
        last_col = get_column_letter(summary_col_start + len(summary_headers) - 1)
        ws.merge_cells(f'A{row}:{last_col}{row}')
        row += 1
        
        # Asset type rows
        for asset_type in asset_types:
            ws[f'A{row}'] = asset_type
            ws[f'A{row}'].font = FONT_NORMAL
            ws[f'A{row}'].alignment = ALIGN_LEFT
            ws[f'A{row}'].border = THIN_BORDER
            
            # Applicability cells (columns B through AE)
            for col_idx in range(2, NUM_STANDARDS + 2):
                col_letter = get_column_letter(col_idx)
                cell = ws[f'{col_letter}{row}']
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_CENTER
                
                # Default to blank (user fills in)
                cell.value = ''
            
            # Summary formulas
            # Required_Standards_Count
            col_letter = get_column_letter(summary_col_start)
            ws[f'{col_letter}{row}'] = f'=COUNTIF(B{row}:{get_column_letter(NUM_STANDARDS+1)}{row},"Required")'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            
            # Recommended_Standards_Count
            col_letter = get_column_letter(summary_col_start + 1)
            ws[f'{col_letter}{row}'] = f'=COUNTIF(B{row}:{get_column_letter(NUM_STANDARDS+1)}{row},"Recommended")'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            
            # Total_Applicable_Standards
            col_letter = get_column_letter(summary_col_start + 2)
            req_col = get_column_letter(summary_col_start)
            rec_col = get_column_letter(summary_col_start + 1)
            ws[f'{col_letter}{row}'] = f'={req_col}{row}+{rec_col}{row}'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            
            # High_Hardening_Burden (Yes if >5 required standards)
            col_letter = get_column_letter(summary_col_start + 3)
            req_col = get_column_letter(summary_col_start)
            ws[f'{col_letter}{row}'] = f'=IF({req_col}{row}>5,"Yes","")'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            
            row += 1
    
    # Add data validation for applicability cells (B3 through last standard column, last asset row)
    last_standard_col = get_column_letter(NUM_STANDARDS + 1)
    last_asset_row = row - 1
    validations = []
    add_data_validation(
        ws,
        f'B3:{last_standard_col}{last_asset_row}',
        DROPDOWNS['applicability'],
        validations=validations
    )
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Conditional formatting for applicability
    for col_idx in range(2, NUM_STANDARDS + 2):
        col_letter = get_column_letter(col_idx)
        cell_range = f'{col_letter}3:{col_letter}{last_asset_row}'

        # Required: Dark green
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Required"'], fill=FILL_DARK_GREEN)
        )

        # Recommended: Light green
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Recommended"'], fill=FILL_LIGHT_GREEN)
        )

        # Optional: Light blue
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Optional"'], fill=FILL_LIGHT_BLUE)
        )

        # Not Applicable: Gray
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Not Applicable"'], fill=FILL_GRAY)
        )

    # Highlight high hardening burden
    burden_col = get_column_letter(summary_col_start + 3)
    ws.conditional_formatting.add(
        f'{burden_col}3:{burden_col}{last_asset_row}',
        CellIsRule(operator='equal', formula=['"Yes"'], fill=FILL_YELLOW)
    )

    # Unlock applicability cells for data entry
    
    # Protect sheet


def create_asset_hardening_assessment(wb: Workbook) -> None:
    """
    Create the Asset Hardening Assessment sheet.
    
    This sheet documents hardening compliance status for individual assets.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Asset Hardening Assessment")
    ws.sheet_view.showGridLines = False
    
    # Define headers
    headers = [
        'Asset ID',
        'Asset Name',
        'Asset Type',
        'Asset Tier',
        'Asset Owner',
        'Location',
        'Operating System',
        'Applicable Standards',
        'Standards Count',
        'Total Controls',
        'Implemented Controls',
        'Partial Controls',
        'Not Implemented Controls',
        'Not Applicable Controls',
        'Compliance Percentage',
        'High Risk Gaps',
        'Medium Risk Gaps',
        'Low Risk Gaps',
        'Active Exceptions',
        'Compensating Controls',
        'Compliance Status',
        'Last Assessment Date',
        'Next Assessment Date',
        'Assessor',
        'Evidence Reference',
        'Remediation Status',
        'Target Compliance Date',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)

    # Freeze panes below header
    ws.freeze_panes = 'A2'

    # Set column widths
    widths = {
        'A': 12,  # Asset_ID
        'B': 30,  # Asset_Name
        'C': 25,  # Asset_Type
        'D': 12,  # Asset_Tier
        'E': 25,  # Asset_Owner
        'F': 20,  # Location
        'G': 25,  # Operating_System
        'H': 25,  # Applicable_Standards
        'I': 15,  # Standards_Count
        'J': 14,  # Total_Controls
        'K': 18,  # Implemented_Controls
        'L': 14,  # Partial_Controls
        'M': 22,  # Not_Implemented_Controls
        'N': 22,  # Not_Applicable_Controls
        'O': 18,  # Compliance_Percentage
        'P': 14,  # High_Risk_Gaps
        'Q': 16,  # Medium_Risk_Gaps
        'R': 14,  # Low_Risk_Gaps
        'S': 16,  # Active_Exceptions
        'T': 22,  # Compensating_Controls
        'U': 20,  # Compliance_Status
        'V': 18,  # Last_Assessment_Date
        'W': 18,  # Next_Assessment_Date
        'X': 25,  # Assessor
        'Y': 20,  # Evidence_Reference
        'Z': 18,  # Remediation_Status
        'AA': 20,  # Target_Compliance_Date
        'AB': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    validations = []
    add_data_validation(ws, 'C2:C101', ALL_ASSET_TYPES, validations=validations)
    add_data_validation(ws, 'D2:D101', DROPDOWNS['asset_tier'], validations=validations)
    # U2:U101 (compliance_status_asset) excluded — column has formula
    add_data_validation(ws, 'Z2:Z101', DROPDOWNS['remediation_status'], validations=validations)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Add formulas for 100 rows
    for row in range(2, 102):
        # Compliance_Percentage formula
        # (Implemented + Partial*0.5) / (Total - Not_Applicable) * 100
        ws[f'O{row}'] = f'=IF(AND(J{row}>0,(J{row}-N{row})>0), ' \
                       f'(K{row}+(L{row}*0.5))/(J{row}-N{row}), "")'
        ws[f'O{row}'].number_format = '0.0%'
        
        # Compliance_Status formula
        ws[f'U{row}'] = (
            f'=IF(O{row}="","", '
            f'IF(AND(O{row}=1, P{row}=0), "Fully Compliant", '
            f'IF(AND(O{row}>=0.95, P{row}=0), "Compliant", '
            f'IF(O{row}>=0.9, "Substantially Compliant", '
            f'IF(O{row}>=0.8, "Partially Compliant", "Non-Compliant")))))'
        )
        
        # Next_Assessment_Date (default 90 days after last assessment)
        ws[f'W{row}'] = f'=IF(V{row}<>"", V{row}+90, "")'
        ws[f'W{row}'].number_format = 'DD.MM.YYYY'
        
        # Standards_Count (count of standards in Applicable_Standards)
        # Simplified: user enters count manually
        ws[f'I{row}'].value = ''
        
        # Set default values
        ws[f'X{row}'].value = '[Security Analyst]'
        ws[f'Z{row}'].value = 'Not Started'
    
    # Conditional formatting
    # Compliance_Status colors
    add_conditional_formatting_status(ws, 'U2:U101', 'U')
    
    # High_Risk_Gaps > 0: Red text
    ws.conditional_formatting.add(
        'P2:P101',
        CellIsRule(operator='greaterThan', formula=['0'], 
                  font=Font(color='C00000', bold=True))
    )
    
    # Critical asset with <100% compliance: Red border
    ws.conditional_formatting.add(
        'A2:AB101',
        FormulaRule(
            formula=['AND(ISNUMBER(SEARCH("Critical",$D2)), $O2<1, $O2<>"")'],
            fill=FILL_LIGHT_RED
        )
    )
    
    # Next_Assessment_Date past due: Red
    ws.conditional_formatting.add(
        'W2:W101',
        CellIsRule(operator='lessThan', formula=['TODAY()'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )
    
    # Next_Assessment_Date within 30 days: Yellow
    ws.conditional_formatting.add(
        'W2:W101',
        CellIsRule(operator='between', formula=['TODAY()', 'TODAY()+30'], fill=FILL_YELLOW)
    )
    
    # Unlock data entry cells
    
    # Protect sheet


def create_control_compliance_detail(wb: Workbook) -> None:
    """
    Create the Control Compliance Detail sheet.
    
    This sheet provides control-level detail for hardening assessments.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Control Compliance Detail")
    ws.sheet_view.showGridLines = False
    
    # Define headers
    headers = [
        'Control ID',
        'Asset ID',
        'Asset Name',
        'Standard ID',
        'Standard Name',
        'Control Number',
        'Control Title',
        'Control Description',
        'Control Category',
        'Control Severity',
        'Implementation Status',
        'Implementation Method',
        'Implementation Evidence',
        'Configuration Setting',
        'Expected Value',
        'Actual Value',
        'Compliance Status',
        'Gap Description',
        'Gap Risk Rating',
        'Remediation Required',
        'Remediation Plan',
        'Remediation Owner',
        'Target Remediation Date',
        'Exception Status',
        'Exception ID',
        'Compensating Control',
        'Last Verified Date',
        'Verified By',
        'Evidence Reference',
        'Verification Method',
        'Next Verification Date',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)

    # Freeze panes below header
    ws.freeze_panes = 'A2'

    # Set column widths
    widths = {
        'A': 12,  # Control_ID
        'B': 12,  # Asset_ID
        'C': 25,  # Asset_Name
        'D': 12,  # Standard_ID
        'E': 30,  # Standard_Name
        'F': 12,  # Control_Number
        'G': 35,  # Control_Title
        'H': 50,  # Control_Description
        'I': 18,  # Control_Category
        'J': 15,  # Control_Severity
        'K': 18,  # Implementation_Status
        'L': 18,  # Implementation_Method
        'M': 40,  # Implementation_Evidence
        'N': 30,  # Configuration_Setting
        'O': 25,  # Expected_Value
        'P': 25,  # Actual_Value
        'Q': 16,  # Compliance_Status
        'R': 40,  # Gap_Description
        'S': 15,  # Gap_Risk_Rating
        'T': 18,  # Remediation_Required
        'U': 40,  # Remediation_Plan
        'V': 25,  # Remediation_Owner
        'W': 20,  # Target_Remediation_Date
        'X': 16,  # Exception_Status
        'Y': 12,  # Exception_ID
        'Z': 40,  # Compensating_Control
        'AA': 18,  # Last_Verified_Date
        'AB': 25,  # Verified_By
        'AC': 20,  # Evidence_Reference
        'AD': 20,  # Verification_Method
        'AE': 20,  # Next_Verification_Date
        'AF': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    validations = []
    add_data_validation(ws, 'I2:I501', DROPDOWNS['control_category'], validations=validations)
    add_data_validation(ws, 'J2:J501', DROPDOWNS['control_severity'], validations=validations)
    add_data_validation(ws, 'K2:K501', DROPDOWNS['implementation_status'], validations=validations)
    add_data_validation(ws, 'L2:L501', DROPDOWNS['implementation_method'], validations=validations)
    # Q2:Q501 (compliance_status_control) excluded — column has formula
    add_data_validation(ws, 'S2:S501', DROPDOWNS['control_severity'], validations=validations)  # Gap_Risk_Rating
    # T2:T501 (yes_no for Remediation_Required) excluded — column has formula
    add_data_validation(ws, 'X2:X501', DROPDOWNS['yes_no'], validations=validations)
    add_data_validation(ws, 'AD2:AD501', DROPDOWNS['verification_method'], validations=validations)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Add formulas for 500 rows
    for row in range(2, 502):
        # Control_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "HC-"&TEXT(ROW()-1,"00000"), "")'
        
        # Compliance_Status formula based on Implementation_Status
        ws[f'Q{row}'] = (
            f'=IF(K{row}="", "", '
            f'IF(ISNUMBER(SEARCH("Implemented",K{row})), "Pass", '
            f'IF(ISNUMBER(SEARCH("Partial",K{row})), "Partial", '
            f'IF(ISNUMBER(SEARCH("Not Applicable",K{row})), "N/A", "Fail"))))'
        )
        
        # Remediation_Required formula
        ws[f'T{row}'] = (
            f'=IF(Q{row}="", "", '
            f'IF(OR(Q{row}="Fail", Q{row}="Partial"), "Yes", "No"))'
        )
        
        # Next_Verification_Date (90 days after last verification)
        ws[f'AE{row}'] = f'=IF(AA{row}<>"", AA{row}+90, "")'
        ws[f'AE{row}'].number_format = 'DD.MM.YYYY'
        
        # Set default values
        ws[f'AB{row}'].value = '[Security Analyst]'
        ws[f'X{row}'].value = 'No'
    
    # Conditional formatting
    # Compliance_Status colors
    add_conditional_formatting_status(ws, 'Q2:Q501', 'Q')
    
    # Critical severity + Fail: Red background, bold
    ws.conditional_formatting.add(
        'A2:AF501',
        FormulaRule(
            formula=['AND(ISNUMBER(SEARCH("Critical",$J2)), $Q2="Fail")'],
            fill=FILL_RED,
            font=Font(bold=True, color='FFFFFF')
        )
    )
    
    # Exception_Status = Yes: Blue background
    ws.conditional_formatting.add(
        'A2:AF501',
        FormulaRule(
            formula=['$X2="Yes"'],
            fill=FILL_LIGHT_BLUE
        )
    )
    
    # Remediation past due
    ws.conditional_formatting.add(
        'W2:W501',
        FormulaRule(
            formula=['AND($T2="Yes", $W2<TODAY(), $W2<>"")'],
            fill=FILL_RED,
            font=Font(color='FFFFFF', bold=True)
        )
    )
    
    # Unlock data entry cells
    
    # Protect sheet


def create_exception_management(wb: Workbook) -> None:
    """
    Create the Exception Management sheet.
    
    This sheet documents and tracks approved deviations from hardening requirements.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Exception Management")
    ws.sheet_view.showGridLines = False

    # Define headers
    headers = [
        'Exception ID',
        'Control ID',
        'Asset ID',
        'Asset Name',
        'Standard ID',
        'Control Number',
        'Control Title',
        'Control Severity',
        'Exception Type',
        'Exception Reason',
        'Business Justification',
        'Risk Assessment',
        'Residual Risk Rating',
        'Compensating Control Required',
        'Compensating Control Description',
        'Compensating Control Effectiveness',
        'Requested By',
        'Request Date',
        'Reviewed By',
        'Review Date',
        'Approved By',
        'Approval Date',
        'Exception Status',
        'Exception Duration',
        'Valid From Date',
        'Valid Until Date',
        'Days Until Expiry',
        'Review Required',
        'Last Review Date',
        'Next Review Date',
        'Audit Trail',
        'Monitoring Required',
        'Monitoring Description',
        'Re Assessment Trigger',
        'Exception Closure Plan',
        'Documentation Reference',
        'Notes',
    ]

    # Add title row
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = "EXCEPTION MANAGEMENT"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Apply headers at row 2
    apply_header_row(ws, 2, headers)

    # Freeze panes below header
    ws.freeze_panes = 'A3'

    # Set column widths
    widths = {
        'A': 12,  # Exception_ID
        'B': 12,  # Control_ID
        'C': 12,  # Asset_ID
        'D': 25,  # Asset_Name
        'E': 12,  # Standard_ID
        'F': 12,  # Control_Number
        'G': 30,  # Control_Title
        'H': 15,  # Control_Severity
        'I': 20,  # Exception_Type
        'J': 40,  # Exception_Reason
        'K': 40,  # Business_Justification
        'L': 40,  # Risk_Assessment
        'M': 18,  # Residual_Risk_Rating
        'N': 25,  # Compensating_Control_Required
        'O': 40,  # Compensating_Control_Description
        'P': 25,  # Compensating_Control_Effectiveness
        'Q': 25,  # Requested_By
        'R': 15,  # Request_Date
        'S': 25,  # Reviewed_By
        'T': 15,  # Review_Date
        'U': 25,  # Approved_By
        'V': 15,  # Approval_Date
        'W': 18,  # Exception_Status
        'X': 18,  # Exception_Duration
        'Y': 15,  # Valid_From_Date
        'Z': 15,  # Valid_Until_Date
        'AA': 16,  # Days_Until_Expiry
        'AB': 18,  # Review_Required
        'AC': 15,  # Last_Review_Date
        'AD': 15,  # Next_Review_Date
        'AE': 40,  # Audit_Trail
        'AF': 18,  # Monitoring_Required
        'AG': 40,  # Monitoring_Description
        'AH': 40,  # Re_Assessment_Trigger
        'AI': 40,  # Exception_Closure_Plan
        'AJ': 30,  # Documentation_Reference
        'AK': 40,  # Notes
    }
    set_column_widths(ws, widths)

    # Add data validation
    validations = []
    add_data_validation(ws, 'I3:I52', DROPDOWNS['exception_type'], validations=validations)
    add_data_validation(ws, 'M3:M52', DROPDOWNS['residual_risk'], validations=validations)
    add_data_validation(ws, 'N3:N52', DROPDOWNS['yes_no'], validations=validations)
    add_data_validation(ws, 'P3:P52', DROPDOWNS['compensating_effectiveness'], validations=validations)
    add_data_validation(ws, 'W3:W52', DROPDOWNS['exception_status'], validations=validations)
    add_data_validation(ws, 'X3:X52', DROPDOWNS['exception_duration'], validations=validations)
    add_data_validation(ws, 'AF3:AF52', DROPDOWNS['yes_no'], validations=validations)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Add formulas for 50 rows
    for row in range(3, 53):
        # Exception_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "EXC-"&TEXT(ROW()-2,"000"), "")'

        # Days_Until_Expiry — manual input (yellow cell)
        ws[f'AA{row}'].fill = FILL_INPUT
        ws[f'AA{row}'].number_format = '0'

        # Review_Required — manual input (yellow cell)
        ws[f'AB{row}'].fill = FILL_INPUT

        # Valid_Until_Date calculation (based on duration)
        # CUSTOMISE: Adjust month calculations if needed
        ws[f'Z{row}'] = (
            f'=IF(AND(Y{row}<>"", X{row}<>""), '
            f'IF(X{row}="3 Months", Y{row}+90, '
            f'IF(X{row}="6 Months", Y{row}+180, '
            f'IF(X{row}="12 Months", Y{row}+365, '
            f'IF(X{row}="24 Months", Y{row}+730, Y{row}+365)))), "")'
        )
        ws[f'Z{row}'].number_format = 'DD.MM.YYYY'

        # Next_Review_Date (midpoint of exception duration)
        ws[f'AD{row}'] = (
            f'=IF(AND(Y{row}<>"", Z{row}<>""), '
            f'Y{row}+((Z{row}-Y{row})/2), "")'
        )
        ws[f'AD{row}'].number_format = 'DD.MM.YYYY'

        # Default values
        ws[f'W{row}'].value = 'Pending Review'
        ws[f'N{row}'].value = 'Yes'
        ws[f'AF{row}'].value = 'Yes'

    # Conditional formatting
    # Exception_Status colors
    add_conditional_formatting_status(ws, 'W3:W52', 'W')

    # Days_Until_Expiry < 30: Yellow
    ws.conditional_formatting.add(
        'AA3:AA52',
        FormulaRule(
            formula=['AND($AA3<30, $AA3>=0)'],
            fill=FILL_YELLOW
        )
    )

    # Days_Until_Expiry < 0 (expired): Red
    ws.conditional_formatting.add(
        'AA3:AA52',
        CellIsRule(operator='lessThan', formula=['0'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )

    # Residual_Risk High/Critical: Red text
    ws.conditional_formatting.add(
        'M3:M52',
        FormulaRule(
            formula=['OR(ISNUMBER(SEARCH("Critical",$M3)), ISNUMBER(SEARCH("High",$M3)))'],
            font=Font(color='C00000', bold=True)
        )
    )

    # Compensating control required but blank description: Red border
    ws.conditional_formatting.add(
        'O3:O52',
        FormulaRule(
            formula=['AND($N3="Yes", $O3="")'],
            border=Border(
                left=Side(style='medium', color='C00000'),
                right=Side(style='medium', color='C00000'),
                top=Side(style='medium', color='C00000'),
                bottom=Side(style='medium', color='C00000')
            )
        )
    )

    # Unlock data entry cells

    # Protect sheet

def create_remediation_tracking(wb: Workbook) -> None:
    """
    Create the Remediation Tracking sheet.
    
    This sheet tracks remediation activities for identified hardening gaps.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Remediation Tracking")
    ws.sheet_view.showGridLines = False

    # Define headers
    headers = [
        'Remediation ID',
        'Gap Type',
        'Control ID',
        'Asset ID',
        'Asset Name',
        'Asset Tier',
        'Standard ID',
        'Control Number',
        'Control Title',
        'Control Severity',
        'Gap Description',
        'Gap Risk Rating',
        'Impact Assessment',
        'Discovery Date',
        'Discovery Method',
        'Remediation Required',
        'Remediation Strategy',
        'Remediation Description',
        'Remediation Owner',
        'Remediation Team',
        'Estimated Effort',
        'Estimated Cost',
        'Dependencies',
        'Remediation Priority',
        'Target Start Date',
        'Target Completion Date',
        'Actual Start Date',
        'Actual Completion Date',
        'Days To Remediate',
        'Days Overdue',
        'Status',
        'Status Notes',
        'Completion Percentage',
        'Blocker Description',
        'Verification Required',
        'Verification Method',
        'Verified By',
        'Verification Date',
        'Verification Result',
        'Re Test Required',
        'Change Request ID',
        'Risk Acceptance ID',
        'Evidence Reference',
        'Lessons Learned',
        'Preventive Action',
        'Closure Date',
        'Approved By',
        'Notes',
    ]

    # Add title row
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = "REMEDIATION TRACKING"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Apply headers at row 2
    apply_header_row(ws, 2, headers)

    # Freeze panes below header
    ws.freeze_panes = 'A3'

    # Set column widths
    widths = {
        'A': 12,  # Remediation_ID
        'B': 18,  # Gap_Type
        'C': 12,  # Control_ID
        'D': 12,  # Asset_ID
        'E': 25,  # Asset_Name
        'F': 12,  # Asset_Tier
        'G': 12,  # Standard_ID
        'H': 12,  # Control_Number
        'I': 30,  # Control_Title
        'J': 15,  # Control_Severity
        'K': 40,  # Gap_Description
        'L': 15,  # Gap_Risk_Rating
        'M': 40,  # Impact_Assessment
        'N': 15,  # Discovery_Date
        'O': 20,  # Discovery_Method
        'P': 18,  # Remediation_Required
        'Q': 20,  # Remediation_Strategy
        'R': 40,  # Remediation_Description
        'S': 25,  # Remediation_Owner
        'T': 25,  # Remediation_Team
        'U': 15,  # Estimated_Effort
        'V': 15,  # Estimated_Cost
        'W': 30,  # Dependencies
        'X': 18,  # Remediation_Priority
        'Y': 15,  # Target_Start_Date
        'Z': 20,  # Target_Completion_Date
        'AA': 15,  # Actual_Start_Date
        'AB': 20,  # Actual_Completion_Date
        'AC': 16,  # Days_To_Remediate
        'AD': 14,  # Days_Overdue
        'AE': 18,  # Status
        'AF': 40,  # Status_Notes
        'AG': 18,  # Completion_Percentage
        'AH': 40,  # Blocker_Description
        'AI': 18,  # Verification_Required
        'AJ': 20,  # Verification_Method
        'AK': 25,  # Verified_By
        'AL': 15,  # Verification_Date
        'AM': 18,  # Verification_Result
        'AN': 15,  # Re_Test_Required
        'AO': 18,  # Change_Request_ID
        'AP': 18,  # Risk_Acceptance_ID
        'AQ': 20,  # Evidence_Reference
        'AR': 40,  # Lessons_Learned
        'AS': 40,  # Preventive_Action
        'AT': 15,  # Closure_Date
        'AU': 25,  # Approved_By
        'AV': 40,  # Notes
    }
    set_column_widths(ws, widths)

    # Add data validation
    validations = []
    add_data_validation(ws, 'B3:B102', DROPDOWNS['gap_type'], validations=validations)
    add_data_validation(ws, 'F3:F102', DROPDOWNS['asset_tier'], validations=validations)
    add_data_validation(ws, 'J3:J102', DROPDOWNS['control_severity'], validations=validations)
    add_data_validation(ws, 'L3:L102', DROPDOWNS['control_severity'], validations=validations)  # Gap_Risk_Rating
    add_data_validation(ws, 'O3:O102', DROPDOWNS['discovery_method'], validations=validations)
    add_data_validation(ws, 'P3:P102', DROPDOWNS['yes_no'], validations=validations)
    add_data_validation(ws, 'Q3:Q102', DROPDOWNS['remediation_strategy'], validations=validations)
    add_data_validation(ws, 'U3:U102', DROPDOWNS['estimated_effort'], validations=validations)
    # X3:X102 (remediation_priority) excluded — column has formula
    add_data_validation(ws, 'AE3:AE102', DROPDOWNS['remediation_tracking_status'], validations=validations)
    add_data_validation(ws, 'AI3:AI102', DROPDOWNS['yes_no'], validations=validations)
    add_data_validation(ws, 'AM3:AM102', DROPDOWNS['verification_result'], validations=validations)
    add_data_validation(ws, 'AN3:AN102', DROPDOWNS['yes_no'], validations=validations)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # Add formulas for 100 rows
    for row in range(3, 103):
        # Remediation_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "REM-"&TEXT(ROW()-2,"000"), "")'

        # Days_To_Remediate
        ws[f'AC{row}'] = f'=IF(AND(AA{row}<>"", AB{row}<>""), AB{row}-AA{row}, "")'
        ws[f'AC{row}'].number_format = '0'

        # Days_Overdue — manual input (yellow cell)
        ws[f'AD{row}'].fill = FILL_INPUT
        ws[f'AD{row}'].number_format = '0'

        # Remediation_Priority based on Gap_Risk_Rating and Asset_Tier
        # CUSTOMISE: Adjust priority logic if needed
        ws[f'X{row}'] = (
            f'=IF(L{row}="", "", '
            f'IF(ISNUMBER(SEARCH("Critical",L{row})), "Critical", '
            f'IF(AND(ISNUMBER(SEARCH("High",L{row})), ISNUMBER(SEARCH("Critical",F{row}))), "Critical", '
            f'IF(ISNUMBER(SEARCH("High",L{row})), "High", '
            f'IF(ISNUMBER(SEARCH("Medium",L{row})), "Medium", "Low")))))'
        )

        # Target_Completion_Date based on priority and discovery date
        # CUSTOMISE: Adjust SLA targets in REMEDIATION_SLA config
        ws[f'Z{row}'] = (
            f'=IF(AND(N{row}<>"", X{row}<>""), '
            f'IF(X{row}="Critical", N{row}+{REMEDIATION_SLA["Critical"]}, '
            f'IF(X{row}="High", N{row}+{REMEDIATION_SLA["High"]}, '
            f'IF(X{row}="Medium", N{row}+{REMEDIATION_SLA["Medium"]}, '
            f'N{row}+{REMEDIATION_SLA["Low"]}))), "")'
        )
        ws[f'Z{row}'].number_format = 'DD.MM.YYYY'

        # Completion_Percentage
        ws[f'AG{row}'].value = 0
        ws[f'AG{row}'].number_format = '0%'

        # Default values
        ws[f'AE{row}'].value = 'Identified'
        ws[f'P{row}'].value = 'Yes'
        ws[f'AI{row}'].value = 'Yes'
        ws[f'AN{row}'].value = 'No'

    # Conditional formatting
    # Status colors
    add_conditional_formatting_status(ws, 'AE3:AE102', 'AE')

    # Days_Overdue > 0: Orange background
    ws.conditional_formatting.add(
        'AD3:AD102',
        CellIsRule(operator='greaterThan', formula=['0'], fill=FILL_ORANGE)
    )

    # Days_Overdue > 7: Red background
    ws.conditional_formatting.add(
        'AD3:AD102',
        CellIsRule(operator='greaterThan', formula=['7'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )

    # Gap_Risk_Rating Critical/High: Red text
    ws.conditional_formatting.add(
        'L3:L102',
        FormulaRule(
            formula=['OR(ISNUMBER(SEARCH("Critical",$L3)), ISNUMBER(SEARCH("High",$L3)))'],
            font=Font(color='C00000', bold=True)
        )
    )

    # Critical asset and not completed: Yellow border
    ws.conditional_formatting.add(
        'A3:AV102',
        FormulaRule(
            formula=['AND(ISNUMBER(SEARCH("Critical",$F3)), NOT(ISNUMBER(SEARCH("Completed",$AE3))))'],
            border=Border(
                left=Side(style='thin', color='FFEB9C'),
                right=Side(style='thin', color='FFEB9C'),
                top=Side(style='thin', color='FFEB9C'),
                bottom=Side(style='thin', color='FFEB9C')
            )
        )
    )

    # Completion_Percentage color scale (handled separately below)
    # Green to Red gradient: 0% = Red, 50% = Yellow, 100% = Green
    from openpyxl.formatting.rule import ColorScaleRule
    color_scale = ColorScaleRule(
        start_type='num', start_value=0, start_color='C00000',
        mid_type='num', mid_value=0.5, mid_color='FFEB9C',
        end_type='num', end_value=1, end_color='C6EFCE'
    )
    ws.conditional_formatting.add('AG3:AG102', color_scale)

    # Unlock data entry cells

    # Protect sheet


def create_compliance_dashboard(wb: Workbook) -> None:
    """
    Create the Gold Standard Summary Dashboard for A.8.9.4 Security Hardening.

    Args:
        wb: Workbook object
    """
    thin = Side(style='thin')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # ── A1: Title bar ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    ws['A1'] = "SECURITY HARDENING — SUMMARY DASHBOARD"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = border_thin
    ws.row_dimensions[1].height = 35

    # ── A2: Subtitle ──────────────────────────────────────────────────────────
    ws.merge_cells('A2:G2')
    ws['A2'] = "Executive summary of hardening standards compliance, control implementation, exception management, and remediation status"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    # ── TABLE 1 Banner ────────────────────────────────────────────────────────
    ws.merge_cells('A4:G4')
    ws['A4'] = "TABLE 1 — HARDENING COMPLIANCE ASSESSMENT BY AREA"
    ws['A4'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].border = border_thin
    ws.row_dimensions[4].height = 20

    # ── TABLE 1 Headers ───────────────────────────────────────────────────────
    t1_headers = ['Assessment Area', 'Total Items', 'Compliant', 'Partial', 'Non-Compliant', 'N/A', 'Compliance %']
    t1_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for col, hdr in zip(t1_cols, t1_headers):
        cell = ws[f'{col}5']
        cell.value = hdr
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    ws.row_dimensions[5].height = 18

    # ── TABLE 1 Data rows ─────────────────────────────────────────────────────
    # Asset Hardening Assessment: col U (col 21) = Compliance_Status
    #   Values: '\u2705 Fully Compliant', '\u2705 Compliant', '\u26A0\uFE0F Substantially Compliant',
    #            '\u26A0\uFE0F Partially Compliant', '\u274C Non-Compliant'
    # Control Compliance Detail: col Q (col 17) = Compliance_Status
    #   Values: '\u2705 Pass', '\u274C Fail', '\u26A0\uFE0F Partial', '➖ N/A'
    # Exception Management: col W (col 23) = Exception_Status
    #   Compliant = '\u2705 Approved' + '\u26A0\uFE0F Conditionally Approved' + '\u2705 Closed'
    #   Non-Compliant = '\u274C Rejected' + '⏰ Expired'
    #   Partial = '\u23F3 Pending Review' + '\u23F3 Under Review'
    # Remediation Tracking: col AE (col 31) = Remediation_Status
    #   Compliant = '\u2705 Completed' + '\u2705 Closed - Fixed' + '➖ Closed - Not Required' + '\u26A0\uFE0F Closed - Exception'
    #   Partial = 'Planning' + '\u2705 Approved' + '\u23F3 In Progress' + '\u23F3 Verification'
    #   Non-Compliant = 'Identified' + '\u274C Blocked' + '\u26A0\uFE0F Deferred'
    t1_data = [
        (
            "Asset Hardening Assessment",
            "=COUNTA('Asset Hardening Assessment'!A:A)-1",
            # Compliant = Fully Compliant + Compliant
            "=COUNTIF('Asset Hardening Assessment'!U:U,\"\u2705 Fully Compliant\")+COUNTIF('Asset Hardening Assessment'!U:U,\"\u2705 Compliant\")",
            # Partial = Substantially + Partially
            "=COUNTIF('Asset Hardening Assessment'!U:U,\"\u26A0\uFE0F Substantially Compliant\")+COUNTIF('Asset Hardening Assessment'!U:U,\"\u26A0\uFE0F Partially Compliant\")",
            # Non-Compliant
            "=COUNTIF('Asset Hardening Assessment'!U:U,\"\u274C Non-Compliant\")",
            # N/A
            "=0",
        ),
        (
            "Control Compliance Detail",
            "=COUNTA('Control Compliance Detail'!B:B)-1",
            # Compliant = Pass (\u2705 Pass)
            "=COUNTIF('Control Compliance Detail'!Q:Q,\"\u2705 Pass\")",
            # Partial = Partial (\u26A0\uFE0F Partial)
            "=COUNTIF('Control Compliance Detail'!Q:Q,\"\u26A0\uFE0F Partial\")",
            # Non-Compliant = Fail (\u274C Fail)
            "=COUNTIF('Control Compliance Detail'!Q:Q,\"\u274C Fail\")",
            # N/A (➖ = \u2796)
            "=COUNTIF('Control Compliance Detail'!Q:Q,\"\u2796 N/A\")",
        ),
        (
            "Exception Management",
            "=COUNTA('Exception Management'!B:B)-2",
            # Compliant = Approved + Conditionally Approved + Closed
            "=COUNTIF('Exception Management'!W:W,\"\u2705 Approved\")+COUNTIF('Exception Management'!W:W,\"\u26A0\uFE0F Conditionally Approved\")+COUNTIF('Exception Management'!W:W,\"\u2705 Closed\")",
            # Partial = Pending + Under Review
            "=COUNTIF('Exception Management'!W:W,\"\u23F3 Pending Review\")+COUNTIF('Exception Management'!W:W,\"\u23F3 Under Review\")",
            # Non-Compliant = Rejected + Expired
            "=COUNTIF('Exception Management'!W:W,\"\u274C Rejected\")+COUNTIF('Exception Management'!W:W,\"\u23F0 Expired\")",
            # N/A
            "=0",
        ),
        (
            "Remediation Tracking",
            "=COUNTA('Remediation Tracking'!B:B)-2",
            # Compliant = Completed + Closed Fixed + Closed Exception + Closed Not Required (➖ = \u2796)
            "=COUNTIF('Remediation Tracking'!AE:AE,\"\u2705 Completed\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u2705 Closed - Fixed\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u26A0\uFE0F Closed - Exception\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u2796 Closed - Not Required\")",
            # Partial = Planning + Approved + In Progress + Verification
            "=COUNTIF('Remediation Tracking'!AE:AE,\"Planning\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u2705 Approved\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u23F3 In Progress\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u23F3 Verification\")",
            # Non-Compliant = Identified + Blocked + Deferred
            "=COUNTIF('Remediation Tracking'!AE:AE,\"Identified\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u274C Blocked\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u26A0\uFE0F Deferred\")",
            # N/A
            "=0",
        ),
    ]

    for i, (area, total, comp, part, noncomp, na) in enumerate(t1_data):
        row = 6 + i
        ws[f'A{row}'] = area
        ws[f'A{row}'].font = Font(name='Calibri', size=11, color='0000FF')
        ws[f'A{row}'].border = border_thin
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

        ws[f'B{row}'] = total
        ws[f'C{row}'] = comp
        ws[f'D{row}'] = part
        ws[f'E{row}'] = noncomp
        ws[f'F{row}'] = na
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].font = Font(name='Calibri', size=11, color='0000FF')
            ws[f'{col}{row}'].border = border_thin
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Compliance % = IF((B-F)=0, 0, C/(B-F))
        ws[f'G{row}'] = f'=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))'
        ws[f'G{row}'].font = Font(name='Calibri', size=11, color='0000FF')
        ws[f'G{row}'].number_format = '0.0%'
        ws[f'G{row}'].border = border_thin
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')

    # ── TABLE 1 TOTAL row ─────────────────────────────────────────────────────
    total_row = 6 + len(t1_data)
    ws[f'A{total_row}'] = "TOTAL"
    ws[f'A{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{total_row}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws[f'A{total_row}'].border = border_thin
    ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{total_row}'] = f'=SUM({col}6:{col}{total_row - 1})'
        ws[f'{col}{total_row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'{col}{total_row}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[f'{col}{total_row}'].border = border_thin
        ws[f'{col}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'G{total_row}'] = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    ws[f'G{total_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'G{total_row}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws[f'G{total_row}'].number_format = '0.0%'
    ws[f'G{total_row}'].border = border_thin
    ws[f'G{total_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # ── TABLE 2 Banner ────────────────────────────────────────────────────────
    t2_start = total_row + 2
    ws.merge_cells(f'A{t2_start}:G{t2_start}')
    ws[f'A{t2_start}'] = "TABLE 2 — KEY HARDENING PERFORMANCE INDICATORS"
    ws[f'A{t2_start}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{t2_start}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{t2_start}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{t2_start}'].border = border_thin

    # TABLE 2 headers
    t2_hrow = t2_start + 1
    for col, hdr in zip(['A', 'B', 'C'], ['KPI', 'Value', 'Target']):
        ws[f'{col}{t2_hrow}'] = hdr
        ws[f'{col}{t2_hrow}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'{col}{t2_hrow}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[f'{col}{t2_hrow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}{t2_hrow}'].border = border_thin
    ws.merge_cells(f'D{t2_hrow}:G{t2_hrow}')
    ws[f'D{t2_hrow}'] = "Notes"
    ws[f'D{t2_hrow}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'D{t2_hrow}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws[f'D{t2_hrow}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'D{t2_hrow}'].border = border_thin

    t2_metrics = [
        ("Assets Fully/Substantially Compliant %",
         f"=IF(B6=0,0,(C6+D6)/B6)",
         f"\u2265{COMPLIANCE_TARGETS['Overall']}%",
         "% of assessed assets at Compliant or Substantially Compliant status"),
        ("Control Pass Rate %",
         "=IF(B7=0,0,C7/(B7-F7))",
         "\u226595%",
         "% of applicable controls with \u2705 Pass status"),
        ("Active Approved Exceptions",
         "=COUNTIF('Exception Management'!W:W,\"\u2705 Approved\")+COUNTIF('Exception Management'!W:W,\"\u26A0\uFE0F Conditionally Approved\")",
         "<5% of controls",
         "Exceptions in approved status — should be minimised"),
        ("Expired Exceptions",
         "=COUNTIF('Exception Management'!W:W,\"\u23F0 Expired\")",
         "0",
         "Exceptions past valid-until date — require immediate renewal or closure"),
        ("Open Remediation Items",
         "=COUNTIF('Remediation Tracking'!AE:AE,\"Identified\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u274C Blocked\")+COUNTIF('Remediation Tracking'!AE:AE,\"\u26A0\uFE0F Deferred\")",
         "0",
         "Remediation items not yet in progress"),
        ("Critical Assets Non-Compliant",
         "=COUNTIFS('Asset Hardening Assessment'!D:D,\"Critical\",'Asset Hardening Assessment'!U:U,\"\u274C Non-Compliant\")",
         "0",
         "Critical tier assets with Non-Compliant status — zero tolerance"),
    ]

    for j, (kpi, val, target, note) in enumerate(t2_metrics):
        r = t2_hrow + 1 + j
        ws[f'A{r}'] = kpi
        ws[f'A{r}'].font = Font(name='Calibri', size=11)
        ws[f'A{r}'].border = border_thin
        ws[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center')

        ws[f'B{r}'] = val
        ws[f'B{r}'].font = Font(name='Calibri', size=11, color='0000FF')
        ws[f'B{r}'].border = border_thin
        ws[f'B{r}'].alignment = Alignment(horizontal='center', vertical='center')
        if '%' in kpi:
            ws[f'B{r}'].number_format = '0.0%'

        ws[f'C{r}'] = target
        ws[f'C{r}'].font = Font(name='Calibri', size=11)
        ws[f'C{r}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'C{r}'].border = border_thin
        ws[f'C{r}'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(f'D{r}:G{r}')
        ws[f'D{r}'] = note
        ws[f'D{r}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
        ws[f'D{r}'].border = border_thin
        ws[f'D{r}'].alignment = Alignment(horizontal='left', vertical='center')

    # ── TABLE 3 Banner ────────────────────────────────────────────────────────
    t3_start = t2_hrow + 1 + len(t2_metrics) + 2
    ws.merge_cells(f'A{t3_start}:G{t3_start}')
    ws[f'A{t3_start}'] = "TABLE 3 — KEY FINDINGS AND RECOMMENDED ACTIONS"
    ws[f'A{t3_start}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{t3_start}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws[f'A{t3_start}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{t3_start}'].border = border_thin

    # TABLE 3 headers
    t3_hrow = t3_start + 1
    for col, hdr in zip(['A', 'B', 'C', 'D'], ['#', 'Finding', 'Priority', 'Recommended Action']):
        ws[f'{col}{t3_hrow}'] = hdr
        ws[f'{col}{t3_hrow}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'{col}{t3_hrow}'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws[f'{col}{t3_hrow}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}{t3_hrow}'].border = border_thin
    ws.merge_cells(f'D{t3_hrow}:G{t3_hrow}')
    ws[f'D{t3_hrow}'].border = border_thin

    t3_findings = [
        ("1", "Asset Hardening Assessment — assets with '\u274C Non-Compliant' status have failed to meet the minimum hardening standard and present unacceptable security risk", "Critical", "Identify all Non-Compliant assets immediately; assign remediation owners; establish 3-day resolution target for Critical assets and 7 days for High tier"),
        ("2", "Control Compliance Detail — controls with '\u274C Fail' status represent specific configuration gaps that directly increase attack surface", "High", "Prioritise Fail controls by severity (Critical first); create remediation plans in Remediation Tracking; verify fixes with automated re-scan"),
        ("3", "Exception Management — expired exceptions ('\u23F0 Expired') represent unapproved deviations from hardening requirements that must be resolved immediately", "High", "Review all expired exceptions within 5 business days; either renew with CISO approval or remediate the underlying gap"),
        ("4", "Remediation Tracking — items in 'Identified' or '\u274C Blocked' status are not progressing toward resolution and may represent systemic organisational barriers", "Medium", "Review stalled remediations with asset owners; resolve blockers or escalate to management; ensure all items have target completion dates and assigned owners"),
    ]

    for k, (num, finding, priority, action) in enumerate(t3_findings):
        r = t3_hrow + 1 + k
        ws[f'A{r}'] = num
        ws[f'A{r}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{r}'].border = border_thin
        ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')

        ws[f'B{r}'] = finding
        ws[f'B{r}'].font = Font(name='Calibri', size=11)
        ws[f'B{r}'].border = border_thin
        ws[f'B{r}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws[f'C{r}'] = priority
        ws[f'C{r}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'C{r}'].border = border_thin
        ws[f'C{r}'].alignment = Alignment(horizontal='center', vertical='center')
        if priority == 'Critical':
            ws[f'C{r}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif priority == 'High':
            ws[f'C{r}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        else:
            ws[f'C{r}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        ws.merge_cells(f'D{r}:G{r}')
        ws[f'D{r}'] = action
        ws[f'D{r}'].font = Font(name='Calibri', size=11)
        ws[f'D{r}'].border = border_thin
        ws[f'D{r}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws.row_dimensions[r].height = 45

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14

    ws.freeze_panes = 'A4'


def create_gap_prioritization(wb: Workbook) -> None:
    """
    Create the Gap Prioritization sheet.
    
    This sheet provides risk-based prioritization of all hardening gaps.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Gap Prioritization")
    ws.sheet_view.showGridLines = False

    # Define headers
    headers = [
        'Priority Rank',
        'Remediation ID',
        'Asset ID',
        'Asset Name',
        'Asset Tier',
        'Control ID',
        'Control Number',
        'Control Title',
        'Standard ID',
        'Control Severity',
        'Gap Risk Rating',
        'Gap Description',
        'Exploitation Likelihood',
        'Impact Assessment',
        'Risk Score',
        'Priority Category',
        'Remediation Owner',
        'Estimated Effort',
        'Target Completion Date',
        'Days Until Target',
        'Status',
        'Dependencies',
        'Quick Win',
        'Related Gaps',
        'Batch Opportunity',
    ]

    # Add title row
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = "GAP PRIORITIZATION"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Apply headers at row 2
    apply_header_row(ws, 2, headers)

    # Freeze panes below header
    ws.freeze_panes = 'A3'

    # Set column widths
    widths = {
        'A': 12,  # Priority_Rank
        'B': 14,  # Remediation_ID
        'C': 12,  # Asset_ID
        'D': 25,  # Asset_Name
        'E': 12,  # Asset_Tier
        'F': 12,  # Control_ID
        'G': 14,  # Control_Number
        'H': 30,  # Control_Title
        'I': 12,  # Standard_ID
        'J': 15,  # Control_Severity
        'K': 15,  # Gap_Risk_Rating
        'L': 40,  # Gap_Description
        'M': 20,  # Exploitation_Likelihood
        'N': 35,  # Impact_Assessment
        'O': 12,  # Risk_Score
        'P': 18,  # Priority_Category
        'Q': 25,  # Remediation_Owner
        'R': 15,  # Estimated_Effort
        'S': 20,  # Target_Completion_Date
        'T': 16,  # Days_Until_Target
        'U': 18,  # Status
        'V': 30,  # Dependencies
        'W': 15,  # Quick_Win
        'X': 25,  # Related_Gaps
        'Y': 25,  # Batch_Opportunity
    }
    set_column_widths(ws, widths)

    # Add data validation
    validations = []
    add_data_validation(ws, 'M3:M102', DROPDOWNS['exploitation_likelihood'], validations=validations)
    for _dv in validations:
        ws.add_data_validation(_dv)

    # This sheet pulls data from Remediation Tracking
    # Both sheets now have title row at 1, headers at 2, data at 3+
    # GP row N maps to RT row N (same offset)
    # Add formulas for 100 rows
    for row in range(3, 103):
        # Pull data from Remediation Tracking
        ws[f'B{row}'] = f"='Remediation Tracking'!A{row}"
        ws[f'C{row}'] = f"='Remediation Tracking'!D{row}"
        ws[f'D{row}'] = f"='Remediation Tracking'!E{row}"
        ws[f'E{row}'] = f"='Remediation Tracking'!F{row}"
        ws[f'F{row}'] = f"='Remediation Tracking'!C{row}"
        ws[f'G{row}'] = f"='Remediation Tracking'!H{row}"
        ws[f'H{row}'] = f"='Remediation Tracking'!I{row}"
        ws[f'I{row}'] = f"='Remediation Tracking'!G{row}"
        ws[f'J{row}'] = f"='Remediation Tracking'!J{row}"
        ws[f'K{row}'] = f"='Remediation Tracking'!L{row}"
        ws[f'L{row}'] = f"='Remediation Tracking'!K{row}"
        ws[f'N{row}'] = f"='Remediation Tracking'!M{row}"
        ws[f'Q{row}'] = f"='Remediation Tracking'!S{row}"
        ws[f'R{row}'] = f"='Remediation Tracking'!U{row}"
        ws[f'S{row}'] = f"='Remediation Tracking'!Z{row}"
        ws[f'U{row}'] = f"='Remediation Tracking'!AE{row}"
        ws[f'V{row}'] = f"='Remediation Tracking'!W{row}"

        # Days_Until_Target — manual input (yellow cell)
        ws[f'T{row}'].fill = FILL_INPUT
        ws[f'T{row}'].number_format = '0'

        # Risk_Score calculation
        # CUSTOMISE: Adjust weights in RISK_WEIGHTS config
        ws[f'O{row}'] = (
            f'=IF(E{row}="", "", '
            f'(IF(ISNUMBER(SEARCH("Critical",E{row})),5,IF(ISNUMBER(SEARCH("High",E{row})),4,IF(ISNUMBER(SEARCH("Medium",E{row})),3,2)))*10) + '
            f'(IF(ISNUMBER(SEARCH("Critical",J{row})),5,IF(ISNUMBER(SEARCH("High",J{row})),4,IF(ISNUMBER(SEARCH("Medium",J{row})),3,2)))*8) + '
            f'(IF(ISNUMBER(SEARCH("Very High",M{row})),5,IF(ISNUMBER(SEARCH("Very Low",M{row})),1,IF(ISNUMBER(SEARCH("High",M{row})),4,IF(ISNUMBER(SEARCH("Medium",M{row})),3,IF(ISNUMBER(SEARCH("Low",M{row})),2,1)))))*6) + '
            f'(IF(T{row}<0, ABS(T{row})/7*2, 0)))'
        )
        ws[f'O{row}'].number_format = '0.0'

        # Priority_Category based on Risk_Score and other factors
        ws[f'P{row}'] = (
            f'=IF(O{row}="", "", '
            f'IF(AND(ISNUMBER(SEARCH("Critical",E{row})), OR(ISNUMBER(SEARCH("Critical",K{row})),ISNUMBER(SEARCH("High",K{row}))), T{row}<0), "P0 - Critical & Urgent", '
            f'IF(AND(ISNUMBER(SEARCH("Critical",E{row})), OR(ISNUMBER(SEARCH("Critical",K{row})),ISNUMBER(SEARCH("High",K{row})))), "P1 - Critical", '
            f'IF(OR(AND(ISNUMBER(SEARCH("High",E{row})),ISNUMBER(SEARCH("High",K{row}))),AND(ISNUMBER(SEARCH("Critical",E{row})),ISNUMBER(SEARCH("Medium",K{row})))), "P2 - High Priority", '
            f'IF(OR(AND(ISNUMBER(SEARCH("Medium",E{row})),ISNUMBER(SEARCH("High",K{row}))),AND(ISNUMBER(SEARCH("High",E{row})),ISNUMBER(SEARCH("Medium",K{row})))), "P3 - Medium Priority", '
            f'"P4 - Low Priority")))))'
        )

        # Quick_Win calculation
        ws[f'W{row}'] = (
            f'=IF(AND(OR(R{row}="<1 hour",R{row}="1-4 hours",R{row}="1 day"), '
            f'OR(K{row}="Medium",K{row}="High",K{row}="Critical")), "Yes - Quick Win", "No")'
        )

        # Priority_Rank (manual sort - user should sort by Risk_Score descending)
        ws[f'A{row}'] = f'=IF(B{row}<>"", ROW()-2, "")'

        # Exploitation_Likelihood - user must fill in
        ws[f'M{row}'].value = 'Medium'  # Default

    # Conditional formatting
    # Priority_Category colors
    ws.conditional_formatting.add(
        'P3:P102',
        CellIsRule(operator='equal', formula=['"P0 - Critical & Urgent"'],
                  fill=FILL_RED,
                  font=Font(color='FFFFFF', bold=True))
    )

    ws.conditional_formatting.add(
        'P3:P102',
        CellIsRule(operator='equal', formula=['"P1 - Critical"'], fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )

    ws.conditional_formatting.add(
        'P3:P102',
        CellIsRule(operator='equal', formula=['"P2 - High Priority"'], fill=FILL_ORANGE)
    )

    ws.conditional_formatting.add(
        'P3:P102',
        CellIsRule(operator='equal', formula=['"P3 - Medium Priority"'], fill=FILL_YELLOW)
    )

    # Quick_Win: Green border
    ws.conditional_formatting.add(
        'A3:Y102',
        FormulaRule(
            formula=['$W3="Yes - Quick Win"'],
            border=Border(
                left=Side(style='medium', color='C6EFCE'),
                right=Side(style='medium', color='C6EFCE'),
                top=Side(style='medium', color='C6EFCE'),
                bottom=Side(style='medium', color='C6EFCE')
            )
        )
    )

    # Days_Until_Target < 0 (overdue): Red
    ws.conditional_formatting.add(
        'T3:T102',
        CellIsRule(operator='lessThan', formula=['0'],
                  fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )

    # Status = Blocked: Gray
    ws.conditional_formatting.add(
        'U3:U102',
        CellIsRule(operator='equal', formula=['"\u274C Blocked"'], fill=FILL_GRAY)
    )

    # Unlock data entry cells

    # Protect sheet


def create_evidence_register(wb: Workbook):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(color="808080")
        ws.cell(row=r, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"
def create_approval_sheet(wb: Workbook):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.9.4 - Security Hardening Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "='Summary Dashboard'!G10"),
        ("Assessment Status:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border_thin
        row += 1

    # Status dropdown on Assessment Status
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border_thin

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Apply borders to all merged cell top-left corners (GS-AS-011)
    _as_thin = Side(style="thin")
    _as_border = Border(left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin)
    for merge_range in ws.merged_cells.ranges:
        tl = ws.cell(merge_range.min_row, merge_range.min_col)
        tl.border = _as_border
    ws.freeze_panes = "A3"
# =============================================================================
# MAIN EXECUTION
# =============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 80)
    logger.info("ISMS Control A.8.9.4 - Security Hardening Assessment")
    logger.info("Workbook Generation Script")
    logger.info("=" * 80)
    logger.info("")
    
    logger.info(f"Organisation: {'[Organisation]'}")
    logger.info(f"Assessment Date: {datetime.now().strftime('%d.%m.%Y')}")
    logger.info(f"Assessor: {'[Security Analyst]'}")
    logger.info(f"Output File: {OUTPUT_FILENAME}")
    logger.info("")
    
    logger.info("Creating workbook...")
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    logger.info("Generating sheets:")
    
    logger.info("  [1/11] Instructions...")
    create_instructions_sheet(wb.create_sheet())
    
    logger.info("  [2/11] Hardening Standard Register...")
    create_hardening_standard_register(wb)
    
    logger.info("  [3/11] Asset Type Hardening Matrix...")
    create_asset_type_hardening_matrix(wb)
    
    logger.info("  [4/11] Asset Hardening Assessment...")
    create_asset_hardening_assessment(wb)
    
    logger.info("  [5/11] Control Compliance Detail...")
    create_control_compliance_detail(wb)
    
    logger.info("  [6/11] Exception Management...")
    create_exception_management(wb)
    
    logger.info("  [7/11] Remediation Tracking...")
    create_remediation_tracking(wb)
    
    logger.info("  [8/11] Gap Prioritization...")
    create_gap_prioritization(wb)
    
    logger.info("  [9/11] Evidence Register...")
    create_evidence_register(wb)
    
    logger.info("  [10/11] Summary Dashboard (Gold Standard)...")
    create_compliance_dashboard(wb)
    
    logger.info("  [11/11] Approval Sign Off...")
    create_approval_sheet(wb)
    
    logger.info("")
    logger.info("Saving workbook...")
    finalize_validations(wb)
    wb.save(output_path)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("✓ Workbook generated successfully!")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Output: {output_path}")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("1. Review the Instructions sheet for comprehensive guidance")
    logger.info("2. Start with Hardening Standard Register - define your applicable standards")
    logger.info("3. Complete Asset Type Hardening Matrix - map standards to asset types")
    logger.info("4. Begin asset assessments in Asset Hardening Assessment")
    logger.info("5. Document control-level detail in Control Compliance Detail")
    logger.info("6. Track gaps in Remediation Tracking")
    logger.info("7. Monitor overall posture via Summary Dashboard")
    logger.info("")
    logger.info("IMPORTANT REMINDERS:")
    logger.info("\u2022 This workbook is a TEMPLATE - customise for your organisation")
    logger.info("\u2022 Define hardening standards appropriate for your context")
    logger.info("\u2022 Collect evidence systematically - every 'Implemented' control needs evidence")
    logger.info("\u2022 Use exceptions sparingly (<5% of controls)")
    logger.info("\u2022 Prioritize remediation based on risk (see Gap Prioritization)")
    logger.info("\u2022 Integrate with A.8.9.3 (Monitoring) for continuous compliance")
    logger.info("")
    logger.info("=" * 80)
    return 0


def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
