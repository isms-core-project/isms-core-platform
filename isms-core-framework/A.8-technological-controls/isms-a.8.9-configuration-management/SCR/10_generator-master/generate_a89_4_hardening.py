#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.9.4 - Security Hardening Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Assessment Domain 4 of 4: Security Hardening and Compliance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific hardening standards, compliance scanning tools,
and gap remediation processes.

Key customization areas:
1. Hardening standards selection (CIS, STIG, vendor guides per your stack)
2. Compliance scanning tools (match your actual deployment)
3. Gap remediation priorities (adapt to your risk appetite)
4. Exception approval authorities (align with organizational roles)
5. Compliance thresholds (based on your security requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
security hardening implementation and compliance verification against ISO
27001:2022 Control A.8.9 requirements.

**Purpose:**
Enables systematic assessment of hardening standard selection, implementation,
compliance verification, gap analysis, and remediation tracking to ensure
secure system configurations across all asset types.

**Assessment Scope:**
- Hardening standards selection and applicability mapping
- Compliance scanning tool deployment and coverage
- Gap analysis and risk assessment
- Hardening exception management and approval
- Implementation evidence and validation
- Compliance reporting and trending
- Remediation tracking and SLA compliance
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and hardening standards
2. Hardening Standards - Standard selection and asset type mapping
3. Compliance Scans - Automated scan results and compliance rates
4. Gap Analysis - Identified gaps, risk assessment, remediation planning
5. Hardening Exceptions - Approved exceptions and compensating controls
6. Implementation Evidence - Pre/post hardening validation
7. Compliance Reports - Quarterly compliance trending and analysis
8. Gap Analysis Summary - Consolidated gaps and remediation requirements
9. Evidence Register - Audit evidence tracking (100+ entries)
10. Approval & Sign-Off - Three-tier approval workflow

**Key Features:**
- Data validation with hardening standard and asset type dropdowns
- Conditional formatting for compliance status and gap severity
- Automated gap identification from scan results
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with compliance scanning tools

**Integration:**
This assessment feeds into the A.8.9.5 Compliance Dashboard, which
consolidates data from all four configuration management assessment domains
for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_4_hardening.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_4_hardening.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_4_hardening.py --date 20250127

Output:
    File: ISMS_IMP_A_8_9_4_Security_Hardening_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize hardening standards by asset type
    2. Document compliance scanning tool deployment
    3. Import compliance scan results
    4. Conduct gap analysis with risk assessment
    5. Document hardening exceptions with compensating controls
    6. Collect pre/post hardening validation evidence
    7. Generate quarterly compliance trending reports
    8. Define remediation actions with timelines and owners
    9. Collect and link audit evidence (scan reports, hardening scripts)
    10. Obtain three-tier stakeholder approvals
    11. Feed results into A.8.9.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Assessment Domain:    4 of 4 (Security Hardening and Compliance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-POL-A.8.9, Section 2.5: Security Hardening & Compliance
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9, Part 1: Technical Standards Reference
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.1: Baseline Configuration Assessment (Domain 1)
    - ISMS-IMP-A.8.9.2: Change Control Assessment (Domain 2)
    - ISMS-IMP-A.8.9.3: Configuration Monitoring Assessment (Domain 3)
    - ISMS-IMP-A.8.9.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.9.4 specification
    - Supports comprehensive security hardening evaluation
    - Integrated with A.8.9.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Hardening Standards:**
Security hardening benchmarks evolve continuously. Review CIS Benchmarks
quarterly for updates, monitor DISA STIG releases, and track vendor security
guide revisions. Deprecated hardening controls must be identified and updated.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of hardening implementation, compliance
scanning, and gap remediation.

**Data Protection:**
Assessment workbooks contain sensitive security details including:
- Hardening compliance scan results with gap details
- Security vulnerability information
- Exception justifications and compensating controls
- System configuration weaknesses

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check hardening compliance rates and gap remediation progress
- Semi-annually: Update hardening standards for benchmark revisions
- Annually: Complete reassessment of all asset types
- Ad-hoc: When new vulnerabilities or hardening guidance published

**Quality Assurance:**
Have security architects, hardening specialists, and compliance officers
validate assessments before using results for reporting or remediation
prioritization.

**Regulatory Alignment:**
Ensure hardening practices align with applicable requirements:
- ISO 27001:2022: Control A.8.9 secure configuration
- CIS Controls v8.1: Control 4 secure configuration requirements
- PCI DSS v4.0.1: System hardening and configuration standards
- Sector-specific: Regulatory hardening requirements
- Internal: Organizational security hardening policies

Customize assessment criteria to include regulatory-specific requirements.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
from datetime import datetime, timedelta
from typing import List, Dict, Tuple

# =============================================================================
# Third-Party Imports
# =============================================================================
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

# =============================================================================
# CONFIGURATION SECTION - CUSTOMIZE FOR YOUR ORGANIZATION
# =============================================================================

# CUSTOMIZE: Assessment metadata
CONFIG = {
    'organization_name': '[Organization]',  # Replace with your organization
    'assessment_date': datetime.now().strftime('%d.%m.%Y'),
    'assessor_name': '[Security Analyst]',  # Replace with actual assessor
    'output_filename': f'ISMS_A_8_9_4_Security_Hardening_Assessment_{datetime.now().strftime("%Y%m%d")}.xlsx',
    'evidence_storage_location': '[File Server Path]',  # Replace with actual path
}

# Document identification
DOCUMENT_ID = "ISMS-IMP-A.8.9.4"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Hardening_{GENERATED_TIMESTAMP}.xlsx"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.8.9: Configuration Management"
WORKBOOK_VERSION = "1.0"

# CUSTOMIZE: Compliance targets for different asset tiers
COMPLIANCE_TARGETS = {
    'Critical': 100,   # Critical assets: 100% compliance required
    'High': 98,        # High assets: 98% compliance target
    'Medium': 95,      # Medium assets: 95% compliance target
    'Low': 90,         # Low assets: 90% compliance target
    'Overall': 95,     # Overall organization target: 95%
}

# CUSTOMIZE: Risk scoring weights (adjust based on your risk model)
RISK_WEIGHTS = {
    'asset_tier': {
        'Critical': 5,
        'High': 4,
        'Medium': 3,
        'Low': 2,
    },
    'control_severity': {
        'Critical': 5,
        'High': 4,
        'Medium': 3,
        'Low': 2,
    },
    'exploitation_likelihood': {
        'Very High': 5,
        'High': 4,
        'Medium': 3,
        'Low': 2,
        'Very Low': 1,
    },
}

# CUSTOMIZE: Remediation SLA targets (days) based on priority
REMEDIATION_SLA = {
    'Critical': 3,      # Critical gaps: 3 days
    'High': 7,          # High gaps: 7 days
    'Medium': 30,       # Medium gaps: 30 days
    'Low': 90,          # Low gaps: 90 days
}

# 43-type asset taxonomy (consistent with A.8.9.1, A.8.9.2, A.8.9.3)
# DO NOT MODIFY - This taxonomy must remain consistent across all A.8.9 assessments
ASSET_TAXONOMY = {
    'Infrastructure': [
        'Physical Server',
        'Virtual Server (VM)',
        'Hypervisor/VM Host',
        'Container Host',
        'Storage Array',
        'Backup Appliance',
        'Database Server',
        'Application Server',
        'Legacy System',
    ],
    'Endpoint': [
        'Corporate Workstation',
        'Corporate Laptop',
        'Executive Device',
        'Mobile Device (Corporate)',
        'Mobile Device (BYOD)',
        'Thin Client',
        'Kiosk/Shared Terminal',
    ],
    'Network Services': [
        'Core Router',
        'Edge Router',
        'L2/L3 Switch',
        'Wireless Access Point',
        'Firewall',
        'IDS/IPS Appliance',
        'Load Balancer',
        'VPN Concentrator',
    ],
    'Applications': [
        'Web Application',
        'API Service',
        'Database Management System',
        'File Server/NAS',
        'Email System',
        'Authentication Service (IAM)',
        'Monitoring/Logging System',
    ],
    'Cloud & Virtual': [
        'IaaS Instance',
        'PaaS Application',
        'SaaS Application Configuration',
        'Container (Docker/K8s)',
        'Serverless Function',
        'Cloud Storage Bucket',
        'Cloud Network Configuration',
    ],
    'IoT & OT': [
        'Building Management System',
        'Physical Security System',
        'Industrial Control System (ICS)',
        'SCADA System',
        'IoT Sensor/Device',
    ],
}

# Flatten asset taxonomy for dropdowns
ALL_ASSET_TYPES = []
for category, types in ASSET_TAXONOMY.items():
    ALL_ASSET_TYPES.extend(types)

# Dropdown value definitions
DROPDOWNS = {
    'asset_tier': ['🔴 Critical', '🟡 High', '🟢 Medium', '⭕ Low'],
    'standard_category': [
        'Industry Benchmark',
        'Government Standard',
        'Regulatory Requirement',
        'Vendor Baseline',
        'Framework Control',
        'Custom Organizational Standard',
    ],
    'compliance_level': ['Level 1', 'Level 2', 'Custom'],
    'mandatory_optional': ['Mandatory', 'Optional'],
    'review_frequency': ['Monthly', 'Quarterly', 'Semi-Annual', 'Annual'],
    'status': ['\u2705 Active', '⏸️ Deprecated', '📝 Planned'],
    'applicability': ['Required', 'Recommended', 'Optional', 'Not Applicable'],
    'compliance_status_asset': [
        '\u2705 Fully Compliant',
        '\u2705 Compliant',
        '\u26A0\uFE0F Substantially Compliant',
        '\u26A0\uFE0F Partially Compliant',
        '\u274C Non-Compliant',
    ],
    'remediation_status': [
        '🔴 Not Started',
        '📝 Planning',
        '⏳ In Progress',
        '🚫 Blocked',
        '\u2705 Completed',
        '➖ Accepted as Exception',
    ],
    'implementation_status': [
        '\u2705 Implemented',
        '\u26A0\uFE0F Partial',
        '\u274C Not Implemented',
        '➖ Not Applicable',
        '📝 Planned',
        '⏳ In Progress',
    ],
    'implementation_method': [
        'Automated Tool',
        'Manual Configuration',
        'Native Feature',
        'Third-Party Tool',
        'Compensating Control',
        'Not Implemented',
    ],
    'control_category': [
        'Access Control',
        'Audit & Logging',
        'Authentication',
        'Network Security',
        'Data Protection',
        'System Hardening',
        'Patch Management',
        'Secure Configuration',
        'Service Minimization',
        'Physical Security',
        'Cryptography',
        'Backup & Recovery',
    ],
    'control_severity': ['🔴 Critical', '🟡 High', '🟢 Medium', '⭕ Low'],
    'compliance_status_control': ['\u2705 Pass', '\u274C Fail', '\u26A0\uFE0F Partial', '➖ N/A'],
    'verification_method': [
        'Configuration Export',
        'Security Tool Report',
        'Manual Inspection',
        'Audit Log Review',
        'Test/Validation',
        'Documentation Review',
    ],
    'yes_no': ['Yes', 'No'],
    'exception_type': [
        'Technical Limitation',
        'Business Requirement',
        'Legacy System',
        'Performance Impact',
        'Vendor Limitation',
        'Cost Prohibitive',
        'Temporary Transition',
        'Regulatory Conflict',
    ],
    'exception_status': [
        '📝 Pending Review',
        '🔍 Under Review',
        '\u2705 Approved',
        '\u26A0\uFE0F Conditionally Approved',
        '\u274C Rejected',
        '⏰ Expired',
        '✔️ Closed',
    ],
    'exception_duration': ['3 Months', '6 Months', '12 Months', '24 Months', 'Indefinite'],
    'residual_risk': ['🔴 Critical', '🟡 High', '🟢 Medium', '⭕ Low'],
    'compensating_effectiveness': ['Full', 'Partial', 'None'],
    'gap_type': [
        'Hardening Gap',
        'Configuration Drift',
        'Vulnerability',
        'Audit Finding',
        'Threat Response',
    ],
    'discovery_method': [
        'Security Assessment',
        'Vulnerability Scan',
        'Configuration Monitoring',
        'Penetration Test',
        'Incident Response',
        'Audit',
        'Threat Intelligence',
    ],
    'remediation_strategy': [
        'Configuration Change',
        'Software Update',
        'Policy Change',
        'Compensating Control',
        'Accept as Exception',
        'Asset Decommission',
        'Defer',
    ],
    'estimated_effort': [
        '<1 hour',
        '1-4 hours',
        '1 day',
        '2-5 days',
        '1-2 weeks',
        '2-4 weeks',
        '>1 month',
    ],
    'remediation_priority': ['🔴 Critical', '🟡 High', '🟢 Medium', '⭕ Low'],
    'remediation_tracking_status': [
        '🔍 Identified',
        '📝 Planning',
        '\u2705 Approved',
        '⏳ In Progress',
        '🔎 Verification',
        '\u2705 Completed',
        '🚫 Blocked',
        '⏸️ Deferred',
        '✔️ Closed - Fixed',
        '➖ Closed - Exception',
        '\u274C Closed - Not Required',
    ],
    'verification_result': ['\u2705 Pass', '\u274C Fail'],
    'evidence_type': [
        'Configuration Export',
        'Screenshot',
        'Security Scan Report',
        'Audit Log Extract',
        'Policy/Procedure Document',
        'Test Results',
        'Attestation',
        'Change Record',
        'Exception Approval',
        'Vendor Documentation',
    ],
    'evidence_source': [
        'Automated Tool',
        'Manual Collection',
        'System Export',
        'Security Tool',
        'Change Management System',
        'Documentation Repository',
        'Email',
        'Meeting Minutes',
    ],
    'collection_method': [
        'API/Script',
        'Command Line',
        'GUI Screenshot',
        'File Export',
        'Report Generation',
        'Manual Documentation',
        'Copy/Paste',
    ],
    'evidence_validity': [
        'Point-in-Time',
        '1 Month',
        '3 Months',
        '6 Months',
        '12 Months',
        'Continuous',
        'Until Changed',
    ],
    'evidence_status': ['\u2705 Active', '⏰ Expired', '🔄 Superseded'],
    'exploitation_likelihood': ['🔴 Very High', '🟡 High', '🟢 Medium', '⭕ Low', '⚪ Very Low'],
}

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

# Fonts
FONT_HEADER = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
FONT_SUBHEADER = Font(name='Calibri', size=12, bold=True)
FONT_NORMAL = Font(name='Calibri', size=11)
FONT_SMALL = Font(name='Calibri', size=10)
FONT_BOLD = Font(name='Calibri', size=11, bold=True)

# Fills
FILL_HEADER = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
FILL_SUBHEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
FILL_LIGHT_BLUE = PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid')
FILL_LIGHT_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_LIGHT_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_LIGHT_RED = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
FILL_GREEN = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_RED = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
FILL_ORANGE = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
FILL_GRAY = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FILL_DARK_GREEN = PatternFill(start_color='375623', end_color='375623', fill_type='solid')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

THICK_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def set_column_widths(ws, widths: Dict[str, float]):
    """
    Set column widths for a worksheet.
    
    Args:
        ws: Worksheet object
        widths: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_header_row(ws, row: int, headers: List[str], start_col: int = 1):
    """
    Apply header formatting to a row.
    
    Args:
        ws: Worksheet object
        row: Row number
        headers: List of header text
        start_col: Starting column (default 1 = A)
    """
    for idx, header in enumerate(headers, start=start_col):
        col_letter = get_column_letter(idx)
        cell = ws[f'{col_letter}{row}']
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, validation_list: List[str], 
                        allow_blank: bool = True, error_title: str = "Invalid Entry",
                        error_message: str = "Please select from the dropdown list."):
    """
    Add dropdown data validation to a cell range.
    
    Args:
        ws: Worksheet object
        cell_range: Cell range (e.g., "B2:B100")
        validation_list: List of valid values
        allow_blank: Allow blank cells
        error_title: Error dialog title
        error_message: Error dialog message
    """
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(validation_list)}"',
        allow_blank=allow_blank,
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_message
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def protect_sheet(ws, password: str = None):
    """
    Protect worksheet while allowing data entry in unprotected cells.
    
    Args:
        ws: Worksheet object
        password: Optional password for protection
    """
    ws.protection.sheet = True
    if password is not None:  # Only set password if provided
        ws.protection.password = password
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False


def unlock_cell_range(ws, cell_range: str):
    """
    Unlock a range of cells for data entry.
    
    Args:
        ws: Worksheet object
        cell_range: Cell range to unlock
    """
    for row in ws[cell_range]:
        for cell in row:
            cell.protection = Protection(locked=False)


def add_conditional_formatting_status(ws, cell_range: str, status_col: str):
    """
    Add conditional formatting for status columns.
    
    Args:
        ws: Worksheet object
        cell_range: Range to apply formatting
        status_col: Column letter of status field
    """
    # Green for positive statuses
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("Fully Compliant",${status_col}2)),ISNUMBER(SEARCH("Compliant",${status_col}2)),'
                    f'ISNUMBER(SEARCH("Completed",${status_col}2)),ISNUMBER(SEARCH("Pass",${status_col}2)),ISNUMBER(SEARCH("Approved",${status_col}2)))'],
            fill=FILL_GREEN
        )
    )
    
    # Yellow for warning statuses
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("Substantially Compliant",${status_col}2)),ISNUMBER(SEARCH("Partial",${status_col}2)),'
                    f'ISNUMBER(SEARCH("In Progress",${status_col}2)),ISNUMBER(SEARCH("Under Review",${status_col}2)))'],
            fill=FILL_YELLOW
        )
    )
    
    # Red for critical statuses
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("Non-Compliant",${status_col}2)),ISNUMBER(SEARCH("Fail",${status_col}2)),'
                    f'ISNUMBER(SEARCH("Blocked",${status_col}2)),ISNUMBER(SEARCH("Rejected",${status_col}2)),ISNUMBER(SEARCH("Expired",${status_col}2)))'],
            fill=FILL_RED
        )
    )


def generate_auto_id(prefix: str, row_num: int, padding: int = 3) -> str:
    """
    Generate auto-incrementing ID with prefix.
    
    Args:
        prefix: ID prefix (e.g., "HS", "HC", "EXC")
        row_num: Row number (1-based)
        padding: Number of digits for padding
        
    Returns:
        Formatted ID string
    """
    return f"{prefix}-{str(row_num).zfill(padding)}"


def calculate_date_offset(base_date_cell: str, offset_days: int, ws_name: str = None) -> str:
    """
    Generate Excel formula for date calculation.
    
    Args:
        base_date_cell: Cell reference for base date
        offset_days: Days to add
        ws_name: Worksheet name if cross-sheet reference
        
    Returns:
        Excel formula string
    """
    if ws_name:
        return f"={ws_name}!{base_date_cell}+{offset_days}"
    return f"={base_date_cell}+{offset_days}"


# =============================================================================
# SHEET GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(wb: Workbook) -> None:
    """
    Create the Instructions sheet with comprehensive guidance.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Instructions", 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100
    
    row = 1

    # Title
    ws[f'A{row}'] = f"{DOCUMENT_ID}  -  Security Hardening Assessment\n{CONTROL_REF}"
    ws[f'A{row}'].font = Font(name='Calibri', size=16, bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells(f'A{row}:B{row+1}')
    ws.row_dimensions[row].height = 40
    row += 2
    row += 1
    
    # Document metadata
    ws[f'A{row}'] = "Document ID:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = DOCUMENT_ID
    row += 1
    
    ws[f'A{row}'] = "Assessment:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "Security Hardening Assessment"
    row += 1
    
    ws[f'A{row}'] = "Version:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = WORKBOOK_VERSION
    row += 1
    
    ws[f'A{row}'] = "Generated:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = datetime.now().strftime("%d.%m.%Y %H:%M")
    row += 2
    
    # Instructions sections
    instructions = [
        ("PURPOSE", [
            "This workbook assesses [Organization]'s compliance with security hardening standards "
            "across all in-scope information assets. Security hardening represents the systematic "
            "reduction of attack surface through implementation of security-focused configuration controls.",
            "",
            "Assessment objectives:",
            "\u2022 Identify which hardening standards apply to which assets",
            "\u2022 Assess implementation of required hardening controls",
            "\u2022 Identify gaps where controls are not implemented",
            "\u2022 Prioritize remediation based on risk",
            "\u2022 Document and track approved exceptions",
            "\u2022 Monitor progress on closing hardening gaps",
        ]),
        
        ("KEY CONCEPTS", [
            "Security Hardening: Systematic reduction of attack surface through security-focused "
            "configuration controls",
            "",
            "Hardening Standard: Documented set of security configuration requirements "
            "(e.g., CIS Benchmark, DISA STIG, vendor baseline)",
            "",
            "Hardening Control: Individual security configuration requirement within a standard",
            "",
            "Implementation Status:",
            "\u2022 Implemented: Control fully implemented and verified",
            "\u2022 Partial: Control partially implemented (60-99% of requirement)",
            "\u2022 Not Implemented: Control not implemented, gap exists",
            "\u2022 Not Applicable: Control does not apply (with justification)",
            "",
            "Exception: Documented deviation from hardening requirement with formal risk acceptance",
        ]),
        
        ("ASSESSMENT WORKFLOW", [
            "1. Define applicable hardening standards in Hardening_Standard_Register",
            "2. Map standards to asset types in Asset_Type_Hardening_Matrix",
            "3. Assess asset-level compliance in Asset_Hardening_Assessment",
            "4. Document control-level detail in Control_Compliance_Detail",
            "5. Manage exceptions in Exception_Management",
            "6. Track remediation in Remediation_Tracking",
            "7. Review Compliance_Dashboard for overall posture",
            "8. Use Gap_Prioritization to sequence remediation",
            "9. Document evidence in Evidence_Register",
            "10. Obtain approvals in Approval_Sign_Off",
        ]),
        
        ("HARDENING STANDARD SELECTION", [
            "Determine applicable standards through risk assessment. Common categories:",
            "",
            "\u2022 Industry Benchmarks: CIS Benchmarks, DISA STIGs, NIST 800-53",
            "\u2022 Regulatory Requirements: PCI-DSS, HIPAA Technical Safeguards, GDPR controls",
            "\u2022 Vendor Baselines: Manufacturer recommended security practices",
            "\u2022 Framework Controls: ISO 27002 technical controls, COBIT",
            "\u2022 Custom Standards: Organization-specific requirements",
            "",
            "Standards should be:",
            "\u2022 Specific enough to assess (not generic)",
            "\u2022 Current version (not outdated)",
            "\u2022 Appropriate for asset type",
            "\u2022 Feasible to implement",
        ]),
        
        ("COMPLIANCE SCORING", [
            f"Overall Compliance = (Implemented + (Partial × 0.5)) / (Total - Not Applicable) × 100%",
            "",
            "Compliance Targets:",
            f"\u2022 Critical Assets: {COMPLIANCE_TARGETS['Critical']}%",
            f"\u2022 High Assets: {COMPLIANCE_TARGETS['High']}%",
            f"\u2022 Medium Assets: {COMPLIANCE_TARGETS['Medium']}%",
            f"\u2022 Low Assets: {COMPLIANCE_TARGETS['Low']}%",
            f"\u2022 Overall Organization: {COMPLIANCE_TARGETS['Overall']}%",
        ]),
        
        ("EXCEPTION MANAGEMENT", [
            "Exceptions are appropriate when:",
            "\u2022 Technical limitation prevents implementation",
            "\u2022 Business requirement conflicts with control",
            "\u2022 Cost exceeds risk reduction benefit",
            "\u2022 Legacy system cannot be modified",
            "",
            "Exception process:",
            "1. Document detailed justification",
            "2. Assess residual risk",
            "3. Identify compensating controls",
            "4. Obtain risk owner approval",
            "5. Set review date (typically 12 months)",
            "",
            "Target: <5% of controls should be exceptions",
        ]),
        
        ("EVIDENCE COLLECTION", [
            "Every 'Implemented' control requires evidence:",
            "",
            "\u2022 Configuration Exports: Baseline vs. current comparison",
            "\u2022 Security Tool Reports: Vulnerability scanners, compliance tools",
            "\u2022 Manual Verification: Screenshots, audit logs",
            "\u2022 Exception Approvals: Risk acceptance documentation",
            "",
            "Evidence must be:",
            "\u2022 Sufficient (adequately demonstrates control)",
            "\u2022 Relevant (directly relates to requirement)",
            "\u2022 Recent (within validity period)",
            "\u2022 Authentic (verifiable source)",
        ]),
        
        ("INTEGRATION POINTS", [
            "This assessment integrates with:",
            "",
            "\u2022 A.8.9.1 (Baseline Configuration): Hardening controls in baselines",
            "\u2022 A.8.9.2 (Change Control): Changes affecting hardening require review",
            "\u2022 A.8.9.3 (Configuration Monitoring): Security drift indicates non-compliance",
            "\u2022 A.8.9.5 (Compliance Dashboard): Aggregates hardening metrics",
            "\u2022 A.5.7 (Threat Intelligence): Emerging threats drive hardening updates",
        ]),
        
        ("ROLES & RESPONSIBILITIES", [
            "Assessment Owner: Coordinates assessments, ensures completion",
            "Asset Owners: Provide configuration data, validate control status",
            "Security Team: Defines standards, assesses compliance",
            "Risk Owner: Approves exceptions, accepts residual risk",
            "Auditor: Verifies evidence, validates compliance claims",
        ]),
        
        ("GETTING STARTED", [
            "1. Start with Hardening_Standard_Register - define your standards",
            "2. Complete Asset_Type_Hardening_Matrix - map standards to asset types",
            "3. Assess assets systematically, starting with Critical tier",
            "4. Document evidence as you go - don't wait until the end",
            "5. Flag exceptions early and initiate approval process",
            "6. Track remediation progress in Remediation_Tracking",
            "7. Review Compliance_Dashboard regularly for status",
        ]),
    ]
    
    for section_title, section_content in instructions:
        # Section header
        ws[f'A{row}'] = section_title
        ws[f'A{row}'].font = FONT_SUBHEADER
        ws[f'A{row}'].fill = FILL_SUBHEADER
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # Section content
        for line in section_content:
            ws[f'B{row}'] = line
            ws[f'B{row}'].font = FONT_NORMAL
            ws[f'B{row}'].alignment = ALIGN_LEFT
            row += 1
        
        row += 1  # Blank line between sections
    
    # Document info at bottom
    row += 2
    ws[f'A{row}'] = 'Document ID:'
    ws[f'B{row}'] = 'ISMS-A.8.9.4-Hardening-Assessment'
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    ws[f'A{row}'] = 'Version:'
    ws[f'B{row}'] = '1.0'
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    ws[f'A{row}'] = 'Date:'
    ws[f'B{row}'] = CONFIG['assessment_date']
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    ws[f'A{row}'] = 'Assessor:'
    ws[f'B{row}'] = CONFIG['assessor_name']
    ws[f'A{row}'].font = FONT_BOLD


def create_hardening_standard_register(wb: Workbook) -> None:
    """
    Create the Hardening_Standard_Register sheet.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Hardening_Standard_Register")
    
    # Define headers
    headers = [
        'Standard_ID',
        'Standard_Name',
        'Standard_Category',
        'Standard_Version',
        'Issuing_Authority',
        'Applicability_Scope',
        'Compliance_Level',
        'Mandatory_Optional',
        'Regulatory_Driver',
        'Control_Count',
        'Implementation_Target',
        'Review_Frequency',
        'Last_Review_Date',
        'Next_Review_Date',
        'Standard_Owner',
        'Documentation_Location',
        'Notes',
        'Status',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Standard_ID
        'B': 35,  # Standard_Name
        'C': 25,  # Standard_Category
        'D': 15,  # Standard_Version
        'E': 25,  # Issuing_Authority
        'F': 30,  # Applicability_Scope
        'G': 15,  # Compliance_Level
        'H': 18,  # Mandatory_Optional
        'I': 25,  # Regulatory_Driver
        'J': 14,  # Control_Count
        'K': 18,  # Implementation_Target
        'L': 18,  # Review_Frequency
        'M': 18,  # Last_Review_Date
        'N': 18,  # Next_Review_Date
        'O': 25,  # Standard_Owner
        'P': 30,  # Documentation_Location
        'Q': 40,  # Notes
        'R': 12,  # Status
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    add_data_validation(ws, 'C2:C31', DROPDOWNS['standard_category'])
    add_data_validation(ws, 'G2:G31', DROPDOWNS['compliance_level'])
    add_data_validation(ws, 'H2:H31', DROPDOWNS['mandatory_optional'])
    add_data_validation(ws, 'L2:L31', DROPDOWNS['review_frequency'])
    add_data_validation(ws, 'R2:R31', DROPDOWNS['status'])
    
    # Add formulas for 30 rows
    for row in range(2, 32):
        # Standard_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "HS-"&TEXT(ROW()-1,"000"), "")'
        ws[f'A{row}'].font = FONT_NORMAL
        ws[f'A{row}'].protection = Protection(locked=True)
        
        # Implementation_Target default 95%
        ws[f'K{row}'].value = 0.95
        ws[f'K{row}'].number_format = '0%'
        
        # Review_Frequency default Quarterly
        ws[f'L{row}'].value = 'Quarterly'
        
        # Status default Active
        ws[f'R{row}'].value = 'Active'
        
        # Mandatory_Optional default Mandatory
        ws[f'H{row}'].value = 'Mandatory'
    
    # Conditional formatting
    # Next_Review_Date past due: Red
    ws.conditional_formatting.add(
        'N2:N31',
        CellIsRule(operator='lessThan', formula=['TODAY()'], fill=FILL_RED)
    )
    
    # Next_Review_Date within 30 days: Yellow
    ws.conditional_formatting.add(
        'N2:N31',
        CellIsRule(operator='between', formula=['TODAY()', 'TODAY()+30'], fill=FILL_YELLOW)
    )
    
    # Status = Deprecated: Gray
    ws.conditional_formatting.add(
        'A2:R31',
        FormulaRule(formula=['$R2="Deprecated"'], fill=FILL_GRAY)
    )
    
    # Mandatory + Target <95%: Orange
    ws.conditional_formatting.add(
        'K2:K31',
        FormulaRule(formula=['AND($H2="Mandatory", $K2<0.95)'], fill=FILL_ORANGE)
    )
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'B2:I31')  # Name through Regulatory_Driver
    unlock_cell_range(ws, 'J2:Q31')  # Control_Count through Notes
    unlock_cell_range(ws, 'R2:R31')  # Status
    
    # Protect sheet
    protect_sheet(ws)

def create_asset_type_hardening_matrix(wb: Workbook) -> None:
    """
    Create the Asset_Type_Hardening_Matrix sheet.
    
    This matrix maps hardening standards (columns) to asset types (rows),
    indicating which standards apply to which asset categories.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Asset_Type_Hardening_Matrix")
    
    # CUSTOMIZE: Number of standards you expect
    # This creates space for 30 standards - adjust if needed
    NUM_STANDARDS = 30
    
    # Row 1: Main header
    ws['A1'] = 'Asset Type'
    ws['A1'].font = FONT_HEADER
    ws['A1'].fill = FILL_HEADER
    ws['A1'].alignment = ALIGN_CENTER
    ws['A1'].border = THIN_BORDER
    
    # Columns B through AE (31 columns): Standard headers
    # These will pull Standard_ID from Hardening_Standard_Register
    for col_idx in range(2, NUM_STANDARDS + 2):  # B to AE
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}1']
        
        # Formula to pull Standard_ID from register
        # Shows "HS-001: Standard Name" format
        cell.value = f'=IF(Hardening_Standard_Register!A{col_idx}<>"", ' \
                    f'Hardening_Standard_Register!A{col_idx}&": "&' \
                    f'LEFT(Hardening_Standard_Register!B{col_idx},20), "")'
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = 18
    
    # Summary columns after standards
    summary_col_start = NUM_STANDARDS + 2
    summary_headers = [
        'Required_Standards_Count',
        'Recommended_Standards_Count',
        'Total_Applicable_Standards',
        'High_Hardening_Burden',
    ]
    
    for idx, header in enumerate(summary_headers):
        col_letter = get_column_letter(summary_col_start + idx)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_SUBHEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = 22
    
    # Set Asset Type column width
    ws.column_dimensions['A'].width = 30
    
    # Row 2 onwards: Asset types (43 rows)
    row = 2
    for category, asset_types in ASSET_TAXONOMY.items():
        # Category header row (merged across all columns)
        ws[f'A{row}'] = f'--- {category.upper()} ---'
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, italic=True)
        ws[f'A{row}'].fill = FILL_LIGHT_BLUE
        ws[f'A{row}'].alignment = ALIGN_LEFT
        
        # Merge category header across all standard columns
        last_col = get_column_letter(summary_col_start + len(summary_headers) - 1)
        ws.merge_cells(f'A{row}:{last_col}{row}')
        row += 1
        
        # Asset type rows
        for asset_type in asset_types:
            ws[f'A{row}'] = asset_type
            ws[f'A{row}'].font = FONT_NORMAL
            ws[f'A{row}'].alignment = ALIGN_LEFT
            ws[f'A{row}'].border = THIN_BORDER
            
            # Applicability cells (columns B through AE)
            for col_idx in range(2, NUM_STANDARDS + 2):
                col_letter = get_column_letter(col_idx)
                cell = ws[f'{col_letter}{row}']
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_CENTER
                
                # Default to blank (user fills in)
                cell.value = ''
            
            # Summary formulas
            # Required_Standards_Count
            col_letter = get_column_letter(summary_col_start)
            ws[f'{col_letter}{row}'] = f'=COUNTIF(B{row}:{get_column_letter(NUM_STANDARDS+1)}{row},"Required")'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            ws[f'{col_letter}{row}'].protection = Protection(locked=True)
            
            # Recommended_Standards_Count
            col_letter = get_column_letter(summary_col_start + 1)
            ws[f'{col_letter}{row}'] = f'=COUNTIF(B{row}:{get_column_letter(NUM_STANDARDS+1)}{row},"Recommended")'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            ws[f'{col_letter}{row}'].protection = Protection(locked=True)
            
            # Total_Applicable_Standards
            col_letter = get_column_letter(summary_col_start + 2)
            req_col = get_column_letter(summary_col_start)
            rec_col = get_column_letter(summary_col_start + 1)
            ws[f'{col_letter}{row}'] = f'={req_col}{row}+{rec_col}{row}'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            ws[f'{col_letter}{row}'].protection = Protection(locked=True)
            
            # High_Hardening_Burden (Yes if >5 required standards)
            col_letter = get_column_letter(summary_col_start + 3)
            req_col = get_column_letter(summary_col_start)
            ws[f'{col_letter}{row}'] = f'=IF({req_col}{row}>5,"Yes","")'
            ws[f'{col_letter}{row}'].font = FONT_NORMAL
            ws[f'{col_letter}{row}'].alignment = ALIGN_CENTER
            ws[f'{col_letter}{row}'].border = THIN_BORDER
            ws[f'{col_letter}{row}'].protection = Protection(locked=True)
            
            row += 1
    
    # Add data validation for applicability cells (B2 through last standard column, last asset row)
    last_standard_col = get_column_letter(NUM_STANDARDS + 1)
    last_asset_row = row - 1
    add_data_validation(
        ws,
        f'B2:{last_standard_col}{last_asset_row}',
        DROPDOWNS['applicability']
    )
    
    # Conditional formatting for applicability
    for col_idx in range(2, NUM_STANDARDS + 2):
        col_letter = get_column_letter(col_idx)
        cell_range = f'{col_letter}2:{col_letter}{last_asset_row}'
        
        # Required: Dark green
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Required"'], fill=FILL_DARK_GREEN)
        )
        
        # Recommended: Light green
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Recommended"'], fill=FILL_LIGHT_GREEN)
        )
        
        # Optional: Light blue
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Optional"'], fill=FILL_LIGHT_BLUE)
        )
        
        # Not Applicable: Gray
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='equal', formula=['"Not Applicable"'], fill=FILL_GRAY)
        )
    
    # Highlight high hardening burden
    burden_col = get_column_letter(summary_col_start + 3)
    ws.conditional_formatting.add(
        f'{burden_col}2:{burden_col}{last_asset_row}',
        CellIsRule(operator='equal', formula=['"Yes"'], fill=FILL_YELLOW)
    )
    
    # Unlock applicability cells for data entry
    unlock_cell_range(ws, f'B2:{last_standard_col}{last_asset_row}')
    
    # Protect sheet
    protect_sheet(ws)


def create_asset_hardening_assessment(wb: Workbook) -> None:
    """
    Create the Asset_Hardening_Assessment sheet.
    
    This sheet documents hardening compliance status for individual assets.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Asset_Hardening_Assessment")
    
    # Define headers
    headers = [
        'Asset_ID',
        'Asset_Name',
        'Asset_Type',
        'Asset_Tier',
        'Asset_Owner',
        'Location',
        'Operating_System',
        'Applicable_Standards',
        'Standards_Count',
        'Total_Controls',
        'Implemented_Controls',
        'Partial_Controls',
        'Not_Implemented_Controls',
        'Not_Applicable_Controls',
        'Compliance_Percentage',
        'High_Risk_Gaps',
        'Medium_Risk_Gaps',
        'Low_Risk_Gaps',
        'Active_Exceptions',
        'Compensating_Controls',
        'Compliance_Status',
        'Last_Assessment_Date',
        'Next_Assessment_Date',
        'Assessor',
        'Evidence_Reference',
        'Remediation_Status',
        'Target_Compliance_Date',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Asset_ID
        'B': 30,  # Asset_Name
        'C': 25,  # Asset_Type
        'D': 12,  # Asset_Tier
        'E': 25,  # Asset_Owner
        'F': 20,  # Location
        'G': 25,  # Operating_System
        'H': 25,  # Applicable_Standards
        'I': 15,  # Standards_Count
        'J': 14,  # Total_Controls
        'K': 18,  # Implemented_Controls
        'L': 14,  # Partial_Controls
        'M': 22,  # Not_Implemented_Controls
        'N': 22,  # Not_Applicable_Controls
        'O': 18,  # Compliance_Percentage
        'P': 14,  # High_Risk_Gaps
        'Q': 16,  # Medium_Risk_Gaps
        'R': 14,  # Low_Risk_Gaps
        'S': 16,  # Active_Exceptions
        'T': 22,  # Compensating_Controls
        'U': 20,  # Compliance_Status
        'V': 18,  # Last_Assessment_Date
        'W': 18,  # Next_Assessment_Date
        'X': 25,  # Assessor
        'Y': 20,  # Evidence_Reference
        'Z': 18,  # Remediation_Status
        'AA': 20,  # Target_Compliance_Date
        'AB': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    add_data_validation(ws, 'C2:C101', ALL_ASSET_TYPES)
    add_data_validation(ws, 'D2:D101', DROPDOWNS['asset_tier'])
    add_data_validation(ws, 'U2:U101', DROPDOWNS['compliance_status_asset'])
    add_data_validation(ws, 'Z2:Z101', DROPDOWNS['remediation_status'])
    
    # Add formulas for 100 rows
    for row in range(2, 102):
        # Compliance_Percentage formula
        # (Implemented + Partial*0.5) / (Total - Not_Applicable) * 100
        ws[f'O{row}'] = f'=IF(AND(J{row}>0,(J{row}-N{row})>0), ' \
                       f'(K{row}+(L{row}*0.5))/(J{row}-N{row}), "")'
        ws[f'O{row}'].number_format = '0.0%'
        ws[f'O{row}'].protection = Protection(locked=True)
        
        # Compliance_Status formula
        ws[f'U{row}'] = (
            f'=IF(O{row}="","", '
            f'IF(AND(O{row}=1, P{row}=0), "Fully Compliant", '
            f'IF(AND(O{row}>=0.95, P{row}=0), "Compliant", '
            f'IF(O{row}>=0.9, "Substantially Compliant", '
            f'IF(O{row}>=0.8, "Partially Compliant", "Non-Compliant")))))'
        )
        ws[f'U{row}'].protection = Protection(locked=True)
        
        # Next_Assessment_Date (default 90 days after last assessment)
        ws[f'W{row}'] = f'=IF(V{row}<>"", V{row}+90, "")'
        ws[f'W{row}'].number_format = 'DD.MM.YYYY'
        ws[f'W{row}'].protection = Protection(locked=True)
        
        # Standards_Count (count of standards in Applicable_Standards)
        # Simplified: user enters count manually
        ws[f'I{row}'].value = ''
        
        # Set default values
        ws[f'X{row}'].value = CONFIG['assessor_name']
        ws[f'Z{row}'].value = 'Not Started'
    
    # Conditional formatting
    # Compliance_Status colors
    add_conditional_formatting_status(ws, 'U2:U101', 'U')
    
    # High_Risk_Gaps > 0: Red text
    ws.conditional_formatting.add(
        'P2:P101',
        CellIsRule(operator='greaterThan', formula=['0'], 
                  font=Font(color='C00000', bold=True))
    )
    
    # Critical asset with <100% compliance: Red border
    ws.conditional_formatting.add(
        'A2:AB101',
        FormulaRule(
            formula=['AND($D2="Critical", $O2<1, $O2<>"")'],
            fill=PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        )
    )
    
    # Next_Assessment_Date past due: Red
    ws.conditional_formatting.add(
        'W2:W101',
        CellIsRule(operator='lessThan', formula=['TODAY()'], fill=FILL_RED)
    )
    
    # Next_Assessment_Date within 30 days: Yellow
    ws.conditional_formatting.add(
        'W2:W101',
        CellIsRule(operator='between', formula=['TODAY()', 'TODAY()+30'], fill=FILL_YELLOW)
    )
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'A2:H101')   # Asset_ID through Applicable_Standards
    unlock_cell_range(ws, 'I2:N101')   # Standards_Count through Not_Applicable_Controls
    unlock_cell_range(ws, 'P2:T101')   # High_Risk_Gaps through Compensating_Controls
    unlock_cell_range(ws, 'V2:V101')   # Last_Assessment_Date
    unlock_cell_range(ws, 'X2:AB101')  # Assessor through Notes
    
    # Protect sheet
    protect_sheet(ws)


def create_control_compliance_detail(wb: Workbook) -> None:
    """
    Create the Control_Compliance_Detail sheet.
    
    This sheet provides control-level detail for hardening assessments.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Control_Compliance_Detail")
    
    # Define headers
    headers = [
        'Control_ID',
        'Asset_ID',
        'Asset_Name',
        'Standard_ID',
        'Standard_Name',
        'Control_Number',
        'Control_Title',
        'Control_Description',
        'Control_Category',
        'Control_Severity',
        'Implementation_Status',
        'Implementation_Method',
        'Implementation_Evidence',
        'Configuration_Setting',
        'Expected_Value',
        'Actual_Value',
        'Compliance_Status',
        'Gap_Description',
        'Gap_Risk_Rating',
        'Remediation_Required',
        'Remediation_Plan',
        'Remediation_Owner',
        'Target_Remediation_Date',
        'Exception_Status',
        'Exception_ID',
        'Compensating_Control',
        'Last_Verified_Date',
        'Verified_By',
        'Evidence_Reference',
        'Verification_Method',
        'Next_Verification_Date',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Control_ID
        'B': 12,  # Asset_ID
        'C': 25,  # Asset_Name
        'D': 12,  # Standard_ID
        'E': 30,  # Standard_Name
        'F': 12,  # Control_Number
        'G': 35,  # Control_Title
        'H': 50,  # Control_Description
        'I': 18,  # Control_Category
        'J': 15,  # Control_Severity
        'K': 18,  # Implementation_Status
        'L': 18,  # Implementation_Method
        'M': 40,  # Implementation_Evidence
        'N': 30,  # Configuration_Setting
        'O': 25,  # Expected_Value
        'P': 25,  # Actual_Value
        'Q': 16,  # Compliance_Status
        'R': 40,  # Gap_Description
        'S': 15,  # Gap_Risk_Rating
        'T': 18,  # Remediation_Required
        'U': 40,  # Remediation_Plan
        'V': 25,  # Remediation_Owner
        'W': 20,  # Target_Remediation_Date
        'X': 16,  # Exception_Status
        'Y': 12,  # Exception_ID
        'Z': 40,  # Compensating_Control
        'AA': 18,  # Last_Verified_Date
        'AB': 25,  # Verified_By
        'AC': 20,  # Evidence_Reference
        'AD': 20,  # Verification_Method
        'AE': 20,  # Next_Verification_Date
        'AF': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    add_data_validation(ws, 'I2:I501', DROPDOWNS['control_category'])
    add_data_validation(ws, 'J2:J501', DROPDOWNS['control_severity'])
    add_data_validation(ws, 'K2:K501', DROPDOWNS['implementation_status'])
    add_data_validation(ws, 'L2:L501', DROPDOWNS['implementation_method'])
    add_data_validation(ws, 'Q2:Q501', DROPDOWNS['compliance_status_control'])
    add_data_validation(ws, 'S2:S501', DROPDOWNS['control_severity'])  # Gap_Risk_Rating
    add_data_validation(ws, 'T2:T501', DROPDOWNS['yes_no'])
    add_data_validation(ws, 'X2:X501', DROPDOWNS['yes_no'])
    add_data_validation(ws, 'AD2:AD501', DROPDOWNS['verification_method'])
    
    # Add formulas for 500 rows
    for row in range(2, 502):
        # Control_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "HC-"&TEXT(ROW()-1,"00000"), "")'
        ws[f'A{row}'].protection = Protection(locked=True)
        
        # Compliance_Status formula based on Implementation_Status
        ws[f'Q{row}'] = (
            f'=IF(K{row}="", "", '
            f'IF(K{row}="Implemented", "Pass", '
            f'IF(K{row}="Partial", "Partial", '
            f'IF(K{row}="Not Applicable", "N/A", "Fail"))))'
        )
        ws[f'Q{row}'].protection = Protection(locked=True)
        
        # Remediation_Required formula
        ws[f'T{row}'] = (
            f'=IF(Q{row}="", "", '
            f'IF(OR(Q{row}="Fail", Q{row}="Partial"), "Yes", "No"))'
        )
        ws[f'T{row}'].protection = Protection(locked=True)
        
        # Next_Verification_Date (90 days after last verification)
        ws[f'AE{row}'] = f'=IF(AA{row}<>"", AA{row}+90, "")'
        ws[f'AE{row}'].number_format = 'DD.MM.YYYY'
        ws[f'AE{row}'].protection = Protection(locked=True)
        
        # Set default values
        ws[f'AB{row}'].value = CONFIG['assessor_name']
        ws[f'X{row}'].value = 'No'
    
    # Conditional formatting
    # Compliance_Status colors
    add_conditional_formatting_status(ws, 'Q2:Q501', 'Q')
    
    # Critical severity + Fail: Red background, bold
    ws.conditional_formatting.add(
        'A2:AF501',
        FormulaRule(
            formula=['AND($J2="Critical", $Q2="Fail")'],
            fill=FILL_RED,
            font=Font(bold=True, color='FFFFFF')
        )
    )
    
    # Exception_Status = Yes: Blue background
    ws.conditional_formatting.add(
        'A2:AF501',
        FormulaRule(
            formula=['$X2="Yes"'],
            fill=PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
        )
    )
    
    # Remediation past due
    ws.conditional_formatting.add(
        'W2:W501',
        FormulaRule(
            formula=['AND($T2="Yes", $W2<TODAY(), $W2<>"")'],
            fill=FILL_RED,
            font=Font(color='FFFFFF', bold=True)
        )
    )
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'B2:P501')   # Asset_ID through Actual_Value
    unlock_cell_range(ws, 'R2:S501')   # Gap_Description, Gap_Risk_Rating
    unlock_cell_range(ws, 'U2:Z501')   # Remediation_Plan through Compensating_Control
    unlock_cell_range(ws, 'AA2:AD501') # Last_Verified_Date through Verification_Method
    unlock_cell_range(ws, 'AF2:AF501') # Notes
    
    # Protect sheet
    protect_sheet(ws)


def create_exception_management(wb: Workbook) -> None:
    """
    Create the Exception_Management sheet.
    
    This sheet documents and tracks approved deviations from hardening requirements.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Exception_Management")
    
    # Define headers
    headers = [
        'Exception_ID',
        'Control_ID',
        'Asset_ID',
        'Asset_Name',
        'Standard_ID',
        'Control_Number',
        'Control_Title',
        'Control_Severity',
        'Exception_Type',
        'Exception_Reason',
        'Business_Justification',
        'Risk_Assessment',
        'Residual_Risk_Rating',
        'Compensating_Control_Required',
        'Compensating_Control_Description',
        'Compensating_Control_Effectiveness',
        'Requested_By',
        'Request_Date',
        'Reviewed_By',
        'Review_Date',
        'Approved_By',
        'Approval_Date',
        'Exception_Status',
        'Exception_Duration',
        'Valid_From_Date',
        'Valid_Until_Date',
        'Days_Until_Expiry',
        'Review_Required',
        'Last_Review_Date',
        'Next_Review_Date',
        'Audit_Trail',
        'Monitoring_Required',
        'Monitoring_Description',
        'Re_Assessment_Trigger',
        'Exception_Closure_Plan',
        'Documentation_Reference',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Exception_ID
        'B': 12,  # Control_ID
        'C': 12,  # Asset_ID
        'D': 25,  # Asset_Name
        'E': 12,  # Standard_ID
        'F': 12,  # Control_Number
        'G': 30,  # Control_Title
        'H': 15,  # Control_Severity
        'I': 20,  # Exception_Type
        'J': 40,  # Exception_Reason
        'K': 40,  # Business_Justification
        'L': 40,  # Risk_Assessment
        'M': 18,  # Residual_Risk_Rating
        'N': 25,  # Compensating_Control_Required
        'O': 40,  # Compensating_Control_Description
        'P': 25,  # Compensating_Control_Effectiveness
        'Q': 25,  # Requested_By
        'R': 15,  # Request_Date
        'S': 25,  # Reviewed_By
        'T': 15,  # Review_Date
        'U': 25,  # Approved_By
        'V': 15,  # Approval_Date
        'W': 18,  # Exception_Status
        'X': 18,  # Exception_Duration
        'Y': 15,  # Valid_From_Date
        'Z': 15,  # Valid_Until_Date
        'AA': 16,  # Days_Until_Expiry
        'AB': 18,  # Review_Required
        'AC': 15,  # Last_Review_Date
        'AD': 15,  # Next_Review_Date
        'AE': 40,  # Audit_Trail
        'AF': 18,  # Monitoring_Required
        'AG': 40,  # Monitoring_Description
        'AH': 40,  # Re_Assessment_Trigger
        'AI': 40,  # Exception_Closure_Plan
        'AJ': 30,  # Documentation_Reference
        'AK': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    add_data_validation(ws, 'I2:I51', DROPDOWNS['exception_type'])
    add_data_validation(ws, 'M2:M51', DROPDOWNS['residual_risk'])
    add_data_validation(ws, 'N2:N51', DROPDOWNS['yes_no'])
    add_data_validation(ws, 'P2:P51', DROPDOWNS['compensating_effectiveness'])
    add_data_validation(ws, 'W2:W51', DROPDOWNS['exception_status'])
    add_data_validation(ws, 'X2:X51', DROPDOWNS['exception_duration'])
    add_data_validation(ws, 'AF2:AF51', DROPDOWNS['yes_no'])
    
    # Add formulas for 50 rows
    for row in range(2, 52):
        # Exception_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "EXC-"&TEXT(ROW()-1,"000"), "")'
        ws[f'A{row}'].protection = Protection(locked=True)
        
        # Days_Until_Expiry
        ws[f'AA{row}'] = f'=IF(Z{row}<>"", Z{row}-TODAY(), "")'
        ws[f'AA{row}'].number_format = '0'
        ws[f'AA{row}'].protection = Protection(locked=True)
        
        # Review_Required formula
        ws[f'AB{row}'] = (
            f'=IF(AA{row}="", "", '
            f'IF(AA{row}<30, "Yes - Expiring Soon", '
            f'IF(AD{row}<TODAY()+30, "Yes - Periodic Review Due", "No")))'
        )
        ws[f'AB{row}'].protection = Protection(locked=True)
        
        # Valid_Until_Date calculation (based on duration)
        # CUSTOMIZE: Adjust month calculations if needed
        ws[f'Z{row}'] = (
            f'=IF(AND(Y{row}<>"", X{row}<>""), '
            f'IF(X{row}="3 Months", Y{row}+90, '
            f'IF(X{row}="6 Months", Y{row}+180, '
            f'IF(X{row}="12 Months", Y{row}+365, '
            f'IF(X{row}="24 Months", Y{row}+730, Y{row}+365)))), "")'
        )
        ws[f'Z{row}'].number_format = 'DD.MM.YYYY'
        ws[f'Z{row}'].protection = Protection(locked=True)
        
        # Next_Review_Date (midpoint of exception duration)
        ws[f'AD{row}'] = (
            f'=IF(AND(Y{row}<>"", Z{row}<>""), '
            f'Y{row}+((Z{row}-Y{row})/2), "")'
        )
        ws[f'AD{row}'].number_format = 'DD.MM.YYYY'
        ws[f'AD{row}'].protection = Protection(locked=True)
        
        # Default values
        ws[f'W{row}'].value = 'Pending Review'
        ws[f'N{row}'].value = 'Yes'
        ws[f'AF{row}'].value = 'Yes'
    
    # Conditional formatting
    # Exception_Status colors
    add_conditional_formatting_status(ws, 'W2:W51', 'W')
    
    # Days_Until_Expiry < 30: Yellow
    ws.conditional_formatting.add(
        'AA2:AA51',
        FormulaRule(
            formula=['AND($AA2<30, $AA2>=0)'],
            fill=FILL_YELLOW
        )
    )
    
    # Days_Until_Expiry < 0 (expired): Red
    ws.conditional_formatting.add(
        'AA2:AA51',
        CellIsRule(operator='lessThan', formula=['0'], fill=FILL_RED)
    )
    
    # Residual_Risk High/Critical: Red text
    ws.conditional_formatting.add(
        'M2:M51',
        FormulaRule(
            formula=['OR($M2="Critical", $M2="High")'],
            font=Font(color='C00000', bold=True)
        )
    )
    
    # Compensating control required but blank description: Red border
    ws.conditional_formatting.add(
        'O2:O51',
        FormulaRule(
            formula=['AND($N2="Yes", $O2="")'],
            border=Border(
                left=Side(style='medium', color='C00000'),
                right=Side(style='medium', color='C00000'),
                top=Side(style='medium', color='C00000'),
                bottom=Side(style='medium', color='C00000')
            )
        )
    )
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'B2:L51')   # Control_ID through Risk_Assessment
    unlock_cell_range(ws, 'M2:P51')   # Residual_Risk_Rating through Effectiveness
    unlock_cell_range(ws, 'Q2:V51')   # Requested_By through Approval_Date
    unlock_cell_range(ws, 'W2:Y51')   # Exception_Status through Valid_From_Date
    unlock_cell_range(ws, 'AC2:AC51') # Last_Review_Date
    unlock_cell_range(ws, 'AE2:AK51') # Audit_Trail through Notes
    
    # Protect sheet
    protect_sheet(ws)

def create_remediation_tracking(wb: Workbook) -> None:
    """
    Create the Remediation_Tracking sheet.
    
    This sheet tracks remediation activities for identified hardening gaps.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Remediation_Tracking")
    
    # Define headers
    headers = [
        'Remediation_ID',
        'Gap_Type',
        'Control_ID',
        'Asset_ID',
        'Asset_Name',
        'Asset_Tier',
        'Standard_ID',
        'Control_Number',
        'Control_Title',
        'Control_Severity',
        'Gap_Description',
        'Gap_Risk_Rating',
        'Impact_Assessment',
        'Discovery_Date',
        'Discovery_Method',
        'Remediation_Required',
        'Remediation_Strategy',
        'Remediation_Description',
        'Remediation_Owner',
        'Remediation_Team',
        'Estimated_Effort',
        'Estimated_Cost',
        'Dependencies',
        'Remediation_Priority',
        'Target_Start_Date',
        'Target_Completion_Date',
        'Actual_Start_Date',
        'Actual_Completion_Date',
        'Days_To_Remediate',
        'Days_Overdue',
        'Status',
        'Status_Notes',
        'Completion_Percentage',
        'Blocker_Description',
        'Verification_Required',
        'Verification_Method',
        'Verified_By',
        'Verification_Date',
        'Verification_Result',
        'Re_Test_Required',
        'Change_Request_ID',
        'Risk_Acceptance_ID',
        'Evidence_Reference',
        'Lessons_Learned',
        'Preventive_Action',
        'Closure_Date',
        'Approved_By',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Remediation_ID
        'B': 18,  # Gap_Type
        'C': 12,  # Control_ID
        'D': 12,  # Asset_ID
        'E': 25,  # Asset_Name
        'F': 12,  # Asset_Tier
        'G': 12,  # Standard_ID
        'H': 12,  # Control_Number
        'I': 30,  # Control_Title
        'J': 15,  # Control_Severity
        'K': 40,  # Gap_Description
        'L': 15,  # Gap_Risk_Rating
        'M': 40,  # Impact_Assessment
        'N': 15,  # Discovery_Date
        'O': 20,  # Discovery_Method
        'P': 18,  # Remediation_Required
        'Q': 20,  # Remediation_Strategy
        'R': 40,  # Remediation_Description
        'S': 25,  # Remediation_Owner
        'T': 25,  # Remediation_Team
        'U': 15,  # Estimated_Effort
        'V': 15,  # Estimated_Cost
        'W': 30,  # Dependencies
        'X': 18,  # Remediation_Priority
        'Y': 15,  # Target_Start_Date
        'Z': 20,  # Target_Completion_Date
        'AA': 15,  # Actual_Start_Date
        'AB': 20,  # Actual_Completion_Date
        'AC': 16,  # Days_To_Remediate
        'AD': 14,  # Days_Overdue
        'AE': 18,  # Status
        'AF': 40,  # Status_Notes
        'AG': 18,  # Completion_Percentage
        'AH': 40,  # Blocker_Description
        'AI': 18,  # Verification_Required
        'AJ': 20,  # Verification_Method
        'AK': 25,  # Verified_By
        'AL': 15,  # Verification_Date
        'AM': 18,  # Verification_Result
        'AN': 15,  # Re_Test_Required
        'AO': 18,  # Change_Request_ID
        'AP': 18,  # Risk_Acceptance_ID
        'AQ': 20,  # Evidence_Reference
        'AR': 40,  # Lessons_Learned
        'AS': 40,  # Preventive_Action
        'AT': 15,  # Closure_Date
        'AU': 25,  # Approved_By
        'AV': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    add_data_validation(ws, 'B2:B101', DROPDOWNS['gap_type'])
    add_data_validation(ws, 'F2:F101', DROPDOWNS['asset_tier'])
    add_data_validation(ws, 'J2:J101', DROPDOWNS['control_severity'])
    add_data_validation(ws, 'L2:L101', DROPDOWNS['control_severity'])  # Gap_Risk_Rating
    add_data_validation(ws, 'O2:O101', DROPDOWNS['discovery_method'])
    add_data_validation(ws, 'P2:P101', DROPDOWNS['yes_no'])
    add_data_validation(ws, 'Q2:Q101', DROPDOWNS['remediation_strategy'])
    add_data_validation(ws, 'U2:U101', DROPDOWNS['estimated_effort'])
    add_data_validation(ws, 'X2:X101', DROPDOWNS['remediation_priority'])
    add_data_validation(ws, 'AE2:AE101', DROPDOWNS['remediation_tracking_status'])
    add_data_validation(ws, 'AI2:AI101', DROPDOWNS['yes_no'])
    add_data_validation(ws, 'AM2:AM101', DROPDOWNS['verification_result'])
    add_data_validation(ws, 'AN2:AN101', DROPDOWNS['yes_no'])
    
    # Add formulas for 100 rows
    for row in range(2, 102):
        # Remediation_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "REM-"&TEXT(ROW()-1,"000"), "")'
        ws[f'A{row}'].protection = Protection(locked=True)
        
        # Days_To_Remediate
        ws[f'AC{row}'] = f'=IF(AND(AA{row}<>"", AB{row}<>""), AB{row}-AA{row}, "")'
        ws[f'AC{row}'].number_format = '0'
        ws[f'AC{row}'].protection = Protection(locked=True)
        
        # Days_Overdue
        ws[f'AD{row}'] = (
            f'=IF(AE{row}="", "", '
            f'IF(OR(AE{row}="Completed", AE{row}="Closed - Fixed", AE{row}="Closed - Exception"), 0, '
            f'IF(AND(Z{row}<>"", TODAY()>Z{row}), TODAY()-Z{row}, 0)))'
        )
        ws[f'AD{row}'].number_format = '0'
        ws[f'AD{row}'].protection = Protection(locked=True)
        
        # Remediation_Priority based on Gap_Risk_Rating and Asset_Tier
        # CUSTOMIZE: Adjust priority logic if needed
        ws[f'X{row}'] = (
            f'=IF(L{row}="", "", '
            f'IF(L{row}="Critical", "Critical", '
            f'IF(AND(L{row}="High", F{row}="Critical"), "Critical", '
            f'IF(L{row}="High", "High", '
            f'IF(L{row}="Medium", "Medium", "Low")))))'
        )
        ws[f'X{row}'].protection = Protection(locked=True)
        
        # Target_Completion_Date based on priority and discovery date
        # CUSTOMIZE: Adjust SLA targets in REMEDIATION_SLA config
        ws[f'Z{row}'] = (
            f'=IF(AND(N{row}<>"", X{row}<>""), '
            f'IF(X{row}="Critical", N{row}+{REMEDIATION_SLA["Critical"]}, '
            f'IF(X{row}="High", N{row}+{REMEDIATION_SLA["High"]}, '
            f'IF(X{row}="Medium", N{row}+{REMEDIATION_SLA["Medium"]}, '
            f'N{row}+{REMEDIATION_SLA["Low"]}))), "")'
        )
        ws[f'Z{row}'].number_format = 'DD.MM.YYYY'
        ws[f'Z{row}'].protection = Protection(locked=True)
        
        # Completion_Percentage
        ws[f'AG{row}'].value = 0
        ws[f'AG{row}'].number_format = '0%'
        
        # Default values
        ws[f'AE{row}'].value = 'Identified'
        ws[f'P{row}'].value = 'Yes'
        ws[f'AI{row}'].value = 'Yes'
        ws[f'AN{row}'].value = 'No'
    
    # Conditional formatting
    # Status colors
    add_conditional_formatting_status(ws, 'AE2:AE101', 'AE')
    
    # Days_Overdue > 0: Orange background
    ws.conditional_formatting.add(
        'AD2:AD101',
        CellIsRule(operator='greaterThan', formula=['0'], fill=FILL_ORANGE)
    )
    
    # Days_Overdue > 7: Red background
    ws.conditional_formatting.add(
        'AD2:AD101',
        CellIsRule(operator='greaterThan', formula=['7'], fill=FILL_RED)
    )
    
    # Gap_Risk_Rating Critical/High: Red text
    ws.conditional_formatting.add(
        'L2:L101',
        FormulaRule(
            formula=['OR($L2="Critical", $L2="High")'],
            font=Font(color='C00000', bold=True)
        )
    )
    
    # Critical asset and not completed: Yellow border
    ws.conditional_formatting.add(
        'A2:AV101',
        FormulaRule(
            formula=['AND($F2="Critical", $AE2<>"Completed")'],
            border=Border(
                left=Side(style='thin', color='FFEB9C'),
                right=Side(style='thin', color='FFEB9C'),
                top=Side(style='thin', color='FFEB9C'),
                bottom=Side(style='thin', color='FFEB9C')
            )
        )
    )
    
    # Completion_Percentage color scale (handled separately below)
    # Green to Red gradient: 0% = Red, 50% = Yellow, 100% = Green
    from openpyxl.formatting.rule import ColorScaleRule
    color_scale = ColorScaleRule(
        start_type='num', start_value=0, start_color='C00000',
        mid_type='num', mid_value=0.5, mid_color='FFEB9C',
        end_type='num', end_value=1, end_color='70AD47'
    )
    ws.conditional_formatting.add('AG2:AG101', color_scale)
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'B2:O101')   # Gap_Type through Discovery_Method
    unlock_cell_range(ws, 'P2:W101')   # Remediation_Required through Dependencies
    unlock_cell_range(ws, 'Y2:AB101')  # Target_Start through Actual_Completion
    unlock_cell_range(ws, 'AE2:AH101') # Status through Blocker_Description
    unlock_cell_range(ws, 'AI2:AV101') # Verification_Required through Notes
    
    # Protect sheet
    protect_sheet(ws)


def create_compliance_dashboard(wb: Workbook) -> None:
    """
    Create the Compliance_Dashboard sheet.
    
    This sheet provides executive-level summary of hardening posture.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Compliance_Dashboard")
    
    # Set column widths for dashboard layout
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    row = 1
    
    # Title
    ws[f'A{row}'] = 'Security Hardening Compliance Dashboard'
    ws[f'A{row}'].font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = ALIGN_CENTER
    row += 1
    
    ws[f'A{row}'] = f'Assessment Date: {CONFIG["assessment_date"]}'
    ws[f'A{row}'].font = FONT_NORMAL
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = ALIGN_CENTER
    row += 2
    
    # Section 1: Overall Compliance Summary
    ws[f'A{row}'] = 'OVERALL COMPLIANCE SUMMARY'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Overall Compliance Percentage
    ws[f'A{row}'] = 'Overall Compliance Percentage'
    ws[f'A{row}'].font = FONT_BOLD
    ws[f'B{row}'] = '=AVERAGE(Asset_Hardening_Assessment!O:O)'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = f'{COMPLIANCE_TARGETS["Overall"]}%'
    row += 1
    
    # Total Assets Assessed
    ws[f'A{row}'] = 'Total Assets Assessed'
    ws[f'B{row}'] = '=COUNTA(Asset_Hardening_Assessment!A:A)-1'
    ws[f'B{row}'].number_format = '0'
    row += 1
    
    # Fully Compliant Assets
    ws[f'A{row}'] = 'Fully Compliant Assets'
    ws[f'B{row}'] = '=COUNTIF(Asset_Hardening_Assessment!U:U,"Fully Compliant")'
    ws[f'B{row}'].number_format = '0'
    ws[f'C{row}'] = 'Percentage'
    ws[f'D{row}'] = f'=B{row}/B{row-1}'
    ws[f'D{row}'].number_format = '0.0%'
    row += 1
    
    # Non-Compliant Assets
    ws[f'A{row}'] = 'Non-Compliant Assets'
    ws[f'B{row}'] = '=COUNTIF(Asset_Hardening_Assessment!U:U,"Non-Compliant")'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # Total Applicable Controls
    ws[f'A{row}'] = 'Total Applicable Controls'
    ws[f'B{row}'] = '=SUM(Asset_Hardening_Assessment!J:J)'
    ws[f'B{row}'].number_format = '#,##0'
    row += 1
    
    # Implemented Controls
    ws[f'A{row}'] = 'Implemented Controls'
    ws[f'B{row}'] = '=SUM(Asset_Hardening_Assessment!K:K)'
    ws[f'B{row}'].number_format = '#,##0'
    row += 1
    
    # Total Gaps
    ws[f'A{row}'] = 'Total Gaps (Not Implemented)'
    ws[f'B{row}'] = '=SUM(Asset_Hardening_Assessment!M:M)'
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # High-Risk Gaps
    ws[f'A{row}'] = 'High-Risk Gaps'
    ws[f'B{row}'] = '=SUM(Asset_Hardening_Assessment!P:P)'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # Active Exceptions
    ws[f'A{row}'] = 'Active Exceptions'
    ws[f'B{row}'] = '=COUNTIF(Exception_Management!W:W,"Approved")'
    ws[f'B{row}'].number_format = '0'
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = f'=B{row-8}*0.05'  # <5% of controls
    ws[f'D{row}'].number_format = '0'
    row += 1
    
    # Open Remediation Items
    ws[f'A{row}'] = 'Open Remediation Items'
    ws[f'B{row}'] = '=COUNTIFS(Remediation_Tracking!AE:AE,"<>Completed",Remediation_Tracking!AE:AE,"<>Closed - Fixed",Remediation_Tracking!AE:AE,"<>Closed - Exception")'
    ws[f'B{row}'].number_format = '0'
    row += 2
    
    # Section 2: Compliance by Asset Tier
    ws[f'A{row}'] = 'COMPLIANCE BY ASSET TIER'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Table headers
    tier_headers = ['Asset Tier', 'Asset Count', 'Avg Compliance %', 'Status']
    for idx, header in enumerate(tier_headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = FONT_BOLD
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BORDER
    row += 1
    
    # Tier rows
    for tier in ['Critical', 'High', 'Medium', 'Low']:
        ws[f'A{row}'] = tier
        ws[f'B{row}'] = f'=COUNTIF(Asset_Hardening_Assessment!D:D,"{tier}")'
        ws[f'B{row}'].number_format = '0'
        ws[f'C{row}'] = f'=AVERAGEIF(Asset_Hardening_Assessment!D:D,"{tier}",Asset_Hardening_Assessment!O:O)'
        ws[f'C{row}'].number_format = '0.0%'
        
        # Status formula (Green/Yellow/Red)
        target = COMPLIANCE_TARGETS[tier] / 100
        ws[f'D{row}'] = (
            f'=IF(C{row}="", "", '
            f'IF(AND(C{row}>={target}, COUNTIFS(Asset_Hardening_Assessment!D:D,"{tier}",'
            f'Asset_Hardening_Assessment!U:U,"Non-Compliant")=0), "Green", '
            f'IF(C{row}>=0.9, "Yellow", "Red")))'
        )
        row += 1
    
    row += 1
    
    # Section 3: Remediation Progress
    ws[f'A{row}'] = 'REMEDIATION PROGRESS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Total Gaps Identified
    ws[f'A{row}'] = 'Total Gaps Identified'
    ws[f'B{row}'] = '=COUNTA(Remediation_Tracking!A:A)-1'
    ws[f'B{row}'].number_format = '0'
    row += 1
    
    # Gaps Closed
    ws[f'A{row}'] = 'Gaps Closed'
    ws[f'B{row}'] = '=COUNTIFS(Remediation_Tracking!AE:AE,"Completed")+COUNTIF(Remediation_Tracking!AE:AE,"Closed - Fixed")'
    ws[f'B{row}'].number_format = '0'
    row += 1
    
    # Gap Closure Rate
    ws[f'A{row}'] = 'Gap Closure Rate'
    ws[f'B{row}'] = f'=IF(B{row-2}>0, B{row-1}/B{row-2}, "")'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '≥80%'
    row += 1
    
    # Mean Time to Remediate (MTTR)
    ws[f'A{row}'] = 'Mean Time to Remediate (MTTR)'
    ws[f'B{row}'] = '=AVERAGE(Remediation_Tracking!AC:AC)'
    ws[f'B{row}'].number_format = '0 "days"'
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '≤14 days'
    row += 1
    
    # Overdue Remediations
    ws[f'A{row}'] = 'Overdue Remediations'
    ws[f'B{row}'] = '=COUNTIF(Remediation_Tracking!AD:AD,">0")'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # Blocked Remediations
    ws[f'A{row}'] = 'Blocked Remediations'
    ws[f'B{row}'] = '=COUNTIF(Remediation_Tracking!AE:AE,"Blocked")'
    ws[f'B{row}'].number_format = '0'
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # Critical Gaps (Open)
    ws[f'A{row}'] = 'Critical Gaps (Open)'
    ws[f'B{row}'] = '=COUNTIFS(Remediation_Tracking!L:L,"Critical",Remediation_Tracking!AE:AE,"<>Completed")'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # High Gaps (Open)
    ws[f'A{row}'] = 'High Gaps (Open)'
    ws[f'B{row}'] = '=COUNTIFS(Remediation_Tracking!L:L,"High",Remediation_Tracking!AE:AE,"<>Completed")'
    ws[f'B{row}'].number_format = '0'
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '≤3'
    row += 2
    
    # Section 4: Exception Analysis
    ws[f'A{row}'] = 'EXCEPTION ANALYSIS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Total Active Exceptions
    ws[f'A{row}'] = 'Total Active Exceptions'
    ws[f'B{row}'] = '=COUNTIF(Exception_Management!W:W,"Approved")'
    ws[f'B{row}'].number_format = '0'
    row += 1
    
    # Exceptions Without Compensating Controls
    ws[f'A{row}'] = 'Exceptions Without Compensating Controls'
    ws[f'B{row}'] = '=COUNTIFS(Exception_Management!N:N,"No",Exception_Management!W:W,"Approved")'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # Exceptions Expiring <30 Days
    ws[f'A{row}'] = 'Exceptions Expiring <30 Days'
    ws[f'B{row}'] = '=COUNTIFS(Exception_Management!AA:AA,"<30",Exception_Management!AA:AA,">=0")'
    ws[f'B{row}'].number_format = '0'
    row += 1
    
    # Expired Exceptions
    ws[f'A{row}'] = 'Expired Exceptions'
    ws[f'B{row}'] = '=COUNTIF(Exception_Management!AA:AA,"<0")'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 1
    
    # High-Risk Exceptions
    ws[f'A{row}'] = 'High-Risk Exceptions (Residual Risk High/Critical)'
    ws[f'B{row}'] = '=COUNTIFS(Exception_Management!M:M,"High",Exception_Management!W:W,"Approved")+COUNTIFS(Exception_Management!M:M,"Critical",Exception_Management!W:W,"Approved")'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 2
    
    # Section 5: Evidence Completeness
    ws[f'A{row}'] = 'EVIDENCE COMPLETENESS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Total Implemented Controls
    ws[f'A{row}'] = 'Total Implemented Controls'
    ws[f'B{row}'] = '=COUNTIF(Control_Compliance_Detail!K:K,"Implemented")'
    ws[f'B{row}'].number_format = '#,##0'
    row += 1
    
    # Controls With Evidence
    ws[f'A{row}'] = 'Controls With Evidence'
    ws[f'B{row}'] = '=COUNTIFS(Control_Compliance_Detail!K:K,"Implemented",Control_Compliance_Detail!AC:AC,"<>")'
    ws[f'B{row}'].number_format = '#,##0'
    row += 1
    
    # Evidence Completeness %
    ws[f'A{row}'] = 'Evidence Completeness %'
    ws[f'B{row}'] = f'=IF(B{row-2}>0, B{row-1}/B{row-2}, "")'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '≥95%'
    row += 1
    
    # Missing Evidence (Critical Assets)
    ws[f'A{row}'] = 'Missing Evidence (Critical Assets)'
    ws[f'B{row}'] = '=COUNTIFS(Control_Compliance_Detail!K:K,"Implemented",Asset_Hardening_Assessment!D:D,"Critical",Control_Compliance_Detail!AC:AC,"")'
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].font = Font(color='C00000', bold=True)
    ws[f'C{row}'] = 'Target'
    ws[f'D{row}'] = '0'
    row += 2
    
    # Note about trend analysis
    ws[f'A{row}'] = 'NOTE: Trend analysis requires 3+ assessment cycles. Maintain historical data for meaningful insights.'
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True)
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = ALIGN_LEFT
    
    # Conditional formatting for dashboard metrics
    # Overall compliance >= 95%: Green
    ws.conditional_formatting.add(
        'B5',
        CellIsRule(operator='greaterThanOrEqual', formula=[str(COMPLIANCE_TARGETS['Overall']/100)], fill=FILL_GREEN)
    )
    
    # Overall compliance 90-94%: Yellow
    ws.conditional_formatting.add(
        'B5',
        CellIsRule(operator='between', formula=['0.90', str((COMPLIANCE_TARGETS['Overall']-1)/100)], fill=FILL_YELLOW)
    )
    
    # Overall compliance <90%: Red
    ws.conditional_formatting.add(
        'B5',
        CellIsRule(operator='lessThan', formula=['0.90'], fill=FILL_RED)
    )


def create_gap_prioritization(wb: Workbook) -> None:
    """
    Create the Gap_Prioritization sheet.
    
    This sheet provides risk-based prioritization of all hardening gaps.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Gap_Prioritization")
    
    # Define headers
    headers = [
        'Priority_Rank',
        'Remediation_ID',
        'Asset_ID',
        'Asset_Name',
        'Asset_Tier',
        'Control_ID',
        'Control_Number',
        'Control_Title',
        'Standard_ID',
        'Control_Severity',
        'Gap_Risk_Rating',
        'Gap_Description',
        'Exploitation_Likelihood',
        'Impact_Assessment',
        'Risk_Score',
        'Priority_Category',
        'Remediation_Owner',
        'Estimated_Effort',
        'Target_Completion_Date',
        'Days_Until_Target',
        'Status',
        'Dependencies',
        'Quick_Win',
        'Related_Gaps',
        'Batch_Opportunity',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Priority_Rank
        'B': 14,  # Remediation_ID
        'C': 12,  # Asset_ID
        'D': 25,  # Asset_Name
        'E': 12,  # Asset_Tier
        'F': 12,  # Control_ID
        'G': 14,  # Control_Number
        'H': 30,  # Control_Title
        'I': 12,  # Standard_ID
        'J': 15,  # Control_Severity
        'K': 15,  # Gap_Risk_Rating
        'L': 40,  # Gap_Description
        'M': 20,  # Exploitation_Likelihood
        'N': 35,  # Impact_Assessment
        'O': 12,  # Risk_Score
        'P': 18,  # Priority_Category
        'Q': 25,  # Remediation_Owner
        'R': 15,  # Estimated_Effort
        'S': 20,  # Target_Completion_Date
        'T': 16,  # Days_Until_Target
        'U': 18,  # Status
        'V': 30,  # Dependencies
        'W': 15,  # Quick_Win
        'X': 25,  # Related_Gaps
        'Y': 25,  # Batch_Opportunity
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    add_data_validation(ws, 'M2:M101', DROPDOWNS['exploitation_likelihood'])
    
    # This sheet pulls data from Remediation_Tracking
    # Add formulas for 100 rows
    for row in range(2, 102):
        # Pull data from Remediation_Tracking
        ws[f'B{row}'] = f'=Remediation_Tracking!A{row}'
        ws[f'C{row}'] = f'=Remediation_Tracking!D{row}'
        ws[f'D{row}'] = f'=Remediation_Tracking!E{row}'
        ws[f'E{row}'] = f'=Remediation_Tracking!F{row}'
        ws[f'F{row}'] = f'=Remediation_Tracking!C{row}'
        ws[f'G{row}'] = f'=Remediation_Tracking!H{row}'
        ws[f'H{row}'] = f'=Remediation_Tracking!I{row}'
        ws[f'I{row}'] = f'=Remediation_Tracking!G{row}'
        ws[f'J{row}'] = f'=Remediation_Tracking!J{row}'
        ws[f'K{row}'] = f'=Remediation_Tracking!L{row}'
        ws[f'L{row}'] = f'=Remediation_Tracking!K{row}'
        ws[f'N{row}'] = f'=Remediation_Tracking!M{row}'
        ws[f'Q{row}'] = f'=Remediation_Tracking!S{row}'
        ws[f'R{row}'] = f'=Remediation_Tracking!U{row}'
        ws[f'S{row}'] = f'=Remediation_Tracking!Z{row}'
        ws[f'U{row}'] = f'=Remediation_Tracking!AE{row}'
        ws[f'V{row}'] = f'=Remediation_Tracking!W{row}'
        
        # Days_Until_Target
        ws[f'T{row}'] = f'=IF(S{row}<>"", S{row}-TODAY(), "")'
        ws[f'T{row}'].number_format = '0'
        
        # Risk_Score calculation
        # CUSTOMIZE: Adjust weights in RISK_WEIGHTS config
        ws[f'O{row}'] = (
            f'=IF(E{row}="", "", '
            f'(IF(E{row}="Critical",5,IF(E{row}="High",4,IF(E{row}="Medium",3,2)))*10) + '
            f'(IF(J{row}="Critical",5,IF(J{row}="High",4,IF(J{row}="Medium",3,2)))*8) + '
            f'(IF(M{row}="Very High",5,IF(M{row}="High",4,IF(M{row}="Medium",3,IF(M{row}="Low",2,1))))*6) + '
            f'(IF(T{row}<0, ABS(T{row})/7*2, 0)))'
        )
        ws[f'O{row}'].number_format = '0.0'
        
        # Priority_Category based on Risk_Score and other factors
        ws[f'P{row}'] = (
            f'=IF(O{row}="", "", '
            f'IF(AND(E{row}="Critical", OR(K{row}="Critical",K{row}="High"), T{row}<0), "P0 - Critical & Urgent", '
            f'IF(AND(E{row}="Critical", OR(K{row}="Critical",K{row}="High")), "P1 - Critical", '
            f'IF(OR(AND(E{row}="High",K{row}="High"),AND(E{row}="Critical",K{row}="Medium")), "P2 - High Priority", '
            f'IF(OR(AND(E{row}="Medium",K{row}="High"),AND(E{row}="High",K{row}="Medium")), "P3 - Medium Priority", '
            f'"P4 - Low Priority")))))'
        )
        
        # Quick_Win calculation
        ws[f'W{row}'] = (
            f'=IF(AND(OR(R{row}="<1 hour",R{row}="1-4 hours",R{row}="1 day"), '
            f'OR(K{row}="Medium",K{row}="High",K{row}="Critical")), "Yes - Quick Win", "No")'
        )
        
        # Priority_Rank (manual sort - user should sort by Risk_Score descending)
        ws[f'A{row}'] = f'=IF(B{row}<>"", ROW()-1, "")'
        
        # Exploitation_Likelihood - user must fill in
        ws[f'M{row}'].value = 'Medium'  # Default
    
    # Conditional formatting
    # Priority_Category colors
    ws.conditional_formatting.add(
        'P2:P101',
        CellIsRule(operator='equal', formula=['"P0 - Critical & Urgent"'], 
                  fill=PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid'),
                  font=Font(color='FFFFFF', bold=True))
    )
    
    ws.conditional_formatting.add(
        'P2:P101',
        CellIsRule(operator='equal', formula=['"P1 - Critical"'], fill=FILL_RED)
    )
    
    ws.conditional_formatting.add(
        'P2:P101',
        CellIsRule(operator='equal', formula=['"P2 - High Priority"'], fill=FILL_ORANGE)
    )
    
    ws.conditional_formatting.add(
        'P2:P101',
        CellIsRule(operator='equal', formula=['"P3 - Medium Priority"'], fill=FILL_YELLOW)
    )
    
    # Quick_Win: Green border
    ws.conditional_formatting.add(
        'A2:Y101',
        FormulaRule(
            formula=['$W2="Yes - Quick Win"'],
            border=Border(
                left=Side(style='medium', color='70AD47'),
                right=Side(style='medium', color='70AD47'),
                top=Side(style='medium', color='70AD47'),
                bottom=Side(style='medium', color='70AD47')
            )
        )
    )
    
    # Days_Until_Target < 0 (overdue): Red
    ws.conditional_formatting.add(
        'T2:T101',
        CellIsRule(operator='lessThan', formula=['0'], 
                  fill=FILL_RED, font=Font(color='FFFFFF', bold=True))
    )
    
    # Status = Blocked: Gray
    ws.conditional_formatting.add(
        'U2:U101',
        CellIsRule(operator='equal', formula=['"Blocked"'], fill=FILL_GRAY)
    )
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'M2:M101')  # Exploitation_Likelihood
    unlock_cell_range(ws, 'X2:Y101')  # Related_Gaps, Batch_Opportunity
    
    # Protect sheet
    protect_sheet(ws)


def create_evidence_register(wb: Workbook) -> None:
    """
    Create the Evidence_Register sheet.
    
    This sheet provides centralized evidence management.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Evidence_Register")
    
    # Define headers
    headers = [
        'Evidence_ID',
        'Evidence_Type',
        'Related_Control_ID',
        'Related_Asset_ID',
        'Related_Exception_ID',
        'Evidence_Description',
        'Evidence_Source',
        'Collection_Method',
        'Collection_Date',
        'Collected_By',
        'File_Name',
        'File_Location',
        'File_Hash',
        'Evidence_Validity_Period',
        'Evidence_Expiry_Date',
        'Review_Required',
        'Evidence_Status',
        'Verification_By',
        'Verification_Date',
        'Audit_Trail',
        'Notes',
    ]
    
    # Apply headers
    apply_header_row(ws, 1, headers)
    
    # Set column widths
    widths = {
        'A': 12,  # Evidence_ID
        'B': 25,  # Evidence_Type
        'C': 14,  # Related_Control_ID
        'D': 14,  # Related_Asset_ID
        'E': 14,  # Related_Exception_ID
        'F': 50,  # Evidence_Description
        'G': 20,  # Evidence_Source
        'H': 20,  # Collection_Method
        'I': 15,  # Collection_Date
        'J': 25,  # Collected_By
        'K': 30,  # File_Name
        'L': 40,  # File_Location
        'M': 65,  # File_Hash
        'N': 20,  # Evidence_Validity_Period
        'O': 18,  # Evidence_Expiry_Date
        'P': 18,  # Review_Required
        'Q': 15,  # Evidence_Status
        'R': 25,  # Verification_By
        'S': 15,  # Verification_Date
        'T': 40,  # Audit_Trail
        'U': 40,  # Notes
    }
    set_column_widths(ws, widths)
    
    # Add data validation
    add_data_validation(ws, 'B2:B101', DROPDOWNS['evidence_type'])
    add_data_validation(ws, 'G2:G101', DROPDOWNS['evidence_source'])
    add_data_validation(ws, 'H2:H101', DROPDOWNS['collection_method'])
    add_data_validation(ws, 'N2:N101', DROPDOWNS['evidence_validity'])
    add_data_validation(ws, 'Q2:Q101', DROPDOWNS['evidence_status'])
    
    # Add formulas for 100 rows
    for row in range(2, 102):
        # Evidence_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"", "EVD-"&TEXT(ROW()-1,"000"), "")'
        ws[f'A{row}'].protection = Protection(locked=True)
        
        # Evidence_Expiry_Date based on Collection_Date and Validity_Period
        ws[f'O{row}'] = (
            f'=IF(AND(I{row}<>"", N{row}<>""), '
            f'IF(N{row}="Point-in-Time", I{row}, '
            f'IF(N{row}="1 Month", I{row}+30, '
            f'IF(N{row}="3 Months", I{row}+90, '
            f'IF(N{row}="6 Months", I{row}+180, '
            f'IF(N{row}="12 Months", I{row}+365, '
            f'IF(N{row}="Continuous", "N/A", I{row})))))), "")'
        )
        ws[f'O{row}'].number_format = 'DD.MM.YYYY'
        ws[f'O{row}'].protection = Protection(locked=True)
        
        # Review_Required
        ws[f'P{row}'] = (
            f'=IF(O{row}="", "", '
            f'IF(Q{row}="Expired", "Yes - Expired", '
            f'IF(AND(O{row}<>"N/A", O{row}<TODAY()+30), "Yes - Expiring Soon", "No")))'
        )
        ws[f'P{row}'].protection = Protection(locked=True)
        
        # Default values
        ws[f'Q{row}'].value = 'Active'
        ws[f'J{row}'].value = CONFIG['assessor_name']
        ws[f'L{row}'].value = CONFIG['evidence_storage_location']
        ws[f'N{row}'].value = '3 Months'
    
    # Conditional formatting
    # Evidence_Status colors
    ws.conditional_formatting.add(
        'Q2:Q101',
        CellIsRule(operator='equal', formula=['"Expired"'], fill=FILL_RED)
    )
    
    ws.conditional_formatting.add(
        'Q2:Q101',
        CellIsRule(operator='equal', formula=['"Superseded"'], fill=FILL_GRAY)
    )
    
    # Review_Required: Yellow
    ws.conditional_formatting.add(
        'P2:P101',
        FormulaRule(formula=['LEFT($P2,3)="Yes"'], fill=FILL_YELLOW)
    )
    
    # Evidence_Expiry_Date within 30 days: Yellow
    ws.conditional_formatting.add(
        'O2:O101',
        FormulaRule(
            formula=['AND($O2<>"", $O2<>"N/A", $O2<TODAY()+30, $O2>=TODAY())'],
            fill=FILL_YELLOW
        )
    )
    
    # Attestation evidence with >3 month validity: Orange border
    ws.conditional_formatting.add(
        'A2:U101',
        FormulaRule(
            formula=['AND($B2="Attestation", OR($N2="6 Months",$N2="12 Months",$N2="Continuous"))'],
            border=Border(
                left=Side(style='thin', color='ED7D31'),
                right=Side(style='thin', color='ED7D31'),
                top=Side(style='thin', color='ED7D31'),
                bottom=Side(style='thin', color='ED7D31')
            )
        )
    )
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'B2:N101')  # Evidence_Type through Evidence_Validity_Period
    unlock_cell_range(ws, 'Q2:U101')  # Evidence_Status through Notes
    
    # Protect sheet
    protect_sheet(ws)


def create_approval_sign_off(wb: Workbook) -> None:
    """
    Create the Approval_Sign_Off sheet.
    
    This sheet documents formal approval of the assessment.
    
    Args:
        wb: Workbook object
    """
    ws = wb.create_sheet("Approval_Sign_Off")
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70
    
    row = 1
    
    # Title
    ws[f'A{row}'] = 'Security Hardening Assessment'
    ws[f'A{row}'].font = Font(name='Calibri', size=16, bold=True)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws[f'A{row}'] = 'Approval & Sign-Off'
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells(f'A{row}:B{row}')
    row += 2
    
    # Section 1: Assessment Summary
    ws[f'A{row}'] = 'ASSESSMENT SUMMARY'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    summary_items = [
        ('Assessment Period', f'{CONFIG["assessment_date"]} (update with actual period)'),
        ('Assessment Scope', 'All in-scope information assets per ISMS scope definition'),
        ('Number of Assets Assessed', '=COUNTA(Asset_Hardening_Assessment!A:A)-1'),
        ('Number of Standards Applied', '=COUNTA(Hardening_Standard_Register!A:A)-1'),
        ('Overall Compliance Percentage', '=Compliance_Dashboard!B5'),
        ('Critical Gaps Identified', '=Compliance_Dashboard!B14'),
        ('High-Risk Gaps Identified', '=Compliance_Dashboard!B13'),
        ('Active Exceptions', '=Compliance_Dashboard!B15'),
    ]
    
    for label, value in summary_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = FONT_BOLD
        if isinstance(value, str) and value.startswith('='):
            ws[f'B{row}'] = value
            if 'Percentage' in label:
                ws[f'B{row}'].number_format = '0.0%'
            elif 'Number' in label or 'Gaps' in label or 'Exceptions' in label:
                ws[f'B{row}'].number_format = '0'
        else:
            ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # Section 2: Key Findings
    ws[f'A{row}'] = 'KEY FINDINGS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws[f'A{row}'] = 'Summary:'
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    
    ws[f'B{row}'] = (
        '[Provide narrative summary of:\n'
        '\u2022 Overall hardening posture assessment\n'
        '\u2022 Significant compliance achievements\n'
        '\u2022 Material gaps or deficiencies identified\n'
        '\u2022 Remediation priorities\n'
        '\u2022 Resource requirements for gap closure\n'
        '\u2022 Recommended actions]'
    )
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 120
    row += 2
    
    # Section 3: Risk Assessment
    ws[f'A{row}'] = 'RISK ASSESSMENT'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    risk_headers = ['Risk Category', 'Status', 'Description']
    for idx, header in enumerate(risk_headers, start=1):
        col_letter = get_column_letter(idx)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = FONT_BOLD
        ws[f'{col_letter}{row}'].fill = FILL_LIGHT_BLUE
        ws[f'{col_letter}{row}'].border = THIN_BORDER
    
    if row == 18:  # Adjust column C width for this section
        ws.column_dimensions['C'].width = 60
    row += 1
    
    risk_items = [
        ('Critical Assets - Non-Compliant', '[Red/Yellow/Green]', '[Description]'),
        ('High-Risk Gaps - Unmitigated', '[Red/Yellow/Green]', '[Description]'),
        ('Exceptions - High Residual Risk', '[Red/Yellow/Green]', '[Description]'),
        ('Remediation - Overdue Items', '[Red/Yellow/Green]', '[Description]'),
        ('Trend - Compliance Direction', '[Red/Yellow/Green]', '[Description]'),
    ]
    
    for category, status, description in risk_items:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = status
        ws[f'C{row}'] = description
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = THIN_BORDER
        row += 1
    
    row += 1
    
    # Section 4: Three-Tier Approval
    ws[f'A{row}'] = 'APPROVAL SIGN-OFF'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Tier 1
    ws[f'A{row}'] = 'Tier 1 - Operational Review'
    ws[f'A{row}'].font = FONT_BOLD
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    tier1_headers = ['Role', 'Name', 'Date', 'Comments']
    for idx, header in enumerate(tier1_headers, start=1):
        col_letter = get_column_letter(idx)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = FONT_BOLD
        ws[f'{col_letter}{row}'].fill = FILL_LIGHT_BLUE
    
    if row >= 24:
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
    row += 1
    
    tier1_roles = [
        ('Security Analyst (Assessor)', CONFIG['assessor_name']),
        ('Security Team Lead', '[Name]'),
    ]
    
    for role, name in tier1_roles:
        ws[f'A{row}'] = role
        ws[f'B{row}'] = name
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'
        ws[f'D{row}'] = ''
        row += 1
    
    row += 1
    
    # Tier 2
    ws[f'A{row}'] = 'Tier 2 - Management Review'
    ws[f'A{row}'].font = FONT_BOLD
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    for idx, header in enumerate(tier1_headers, start=1):
        col_letter = get_column_letter(idx)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = FONT_BOLD
        ws[f'{col_letter}{row}'].fill = FILL_LIGHT_BLUE
    row += 1
    
    tier2_roles = [
        ('Security Manager', '[Name]'),
        ('IT Manager', '[Name]'),
    ]
    
    for role, name in tier2_roles:
        ws[f'A{row}'] = role
        ws[f'B{row}'] = name
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'
        ws[f'D{row}'] = ''
        row += 1
    
    row += 1
    
    # Tier 3
    ws[f'A{row}'] = 'Tier 3 - Executive Acceptance'
    ws[f'A{row}'].font = FONT_BOLD
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    for idx, header in enumerate(tier1_headers, start=1):
        col_letter = get_column_letter(idx)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = FONT_BOLD
        ws[f'{col_letter}{row}'].fill = FILL_LIGHT_BLUE
    row += 1
    
    ws[f'A{row}'] = 'CISO / CIO'
    ws[f'B{row}'] = '[Name]'
    ws[f'C{row}'].number_format = 'DD.MM.YYYY'
    ws[f'D{row}'] = ''
    row += 2
    
    # Section 5: Next Steps
    ws[f'A{row}'] = 'NEXT STEPS'
    ws[f'A{row}'].font = FONT_SUBHEADER
    ws[f'A{row}'].fill = FILL_SUBHEADER
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    next_steps = [
        '☐ Remediation plan approved and resourced',
        '☐ High-risk gaps scheduled for immediate remediation',
        '☐ Exception renewals/reviews scheduled',
        '☐ Next assessment scheduled for [Date]',
        '☐ Evidence collection automated where feasible',
        '☐ Integration with A.8.9.3 (Monitoring) confirmed',
    ]
    
    for step in next_steps:
        ws[f'B{row}'] = step
        ws[f'B{row}'].font = FONT_NORMAL
        row += 1
    
    # Unlock data entry cells
    unlock_cell_range(ws, 'B6:B13')    # Assessment summary editable fields
    unlock_cell_range(ws, 'B16:B16')   # Key Findings
    unlock_cell_range(ws, 'B19:C23')   # Risk Assessment
    unlock_cell_range(ws, 'B26:D27')   # Tier 1 approvals
    unlock_cell_range(ws, 'B31:D32')   # Tier 2 approvals
    unlock_cell_range(ws, 'B36:D36')   # Tier 3 approval
    unlock_cell_range(ws, 'B40:B45')   # Next Steps checkboxes


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """
    Main execution function to generate the Security Hardening Assessment workbook.
    """
    logger.info("=" * 80)
    logger.info("ISMS Control A.8.9.4 - Security Hardening Assessment")
    logger.info("Workbook Generation Script")
    logger.info("=" * 80)
    logger.info("")
    
    logger.info(f"Organization: {CONFIG['organization_name']}")
    logger.info(f"Assessment Date: {CONFIG['assessment_date']}")
    logger.info(f"Assessor: {CONFIG['assessor_name']}")
    logger.info(f"Output File: {CONFIG['output_filename']}")
    logger.info("")
    
    logger.info("Creating workbook...")
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    logger.info("Generating sheets:")
    
    logger.info("  [1/11] Instructions...")
    create_instructions_sheet(wb)
    
    logger.info("  [2/11] Hardening_Standard_Register...")
    create_hardening_standard_register(wb)
    
    logger.info("  [3/11] Asset_Type_Hardening_Matrix...")
    create_asset_type_hardening_matrix(wb)
    
    logger.info("  [4/11] Asset_Hardening_Assessment...")
    create_asset_hardening_assessment(wb)
    
    logger.info("  [5/11] Control_Compliance_Detail...")
    create_control_compliance_detail(wb)
    
    logger.info("  [6/11] Exception_Management...")
    create_exception_management(wb)
    
    logger.info("  [7/11] Remediation_Tracking...")
    create_remediation_tracking(wb)
    
    logger.info("  [8/11] Compliance_Dashboard...")
    create_compliance_dashboard(wb)
    
    logger.info("  [9/11] Gap_Prioritization...")
    create_gap_prioritization(wb)
    
    logger.info("  [10/11] Evidence_Register...")
    create_evidence_register(wb)
    
    logger.info("  [11/11] Approval_Sign_Off...")
    create_approval_sign_off(wb)
    
    logger.info("")
    logger.info("Saving workbook...")
    wb.save(CONFIG['output_filename'])
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("✓ Workbook generated successfully!")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Output: {CONFIG['output_filename']}")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("1. Review the Instructions sheet for comprehensive guidance")
    logger.info("2. Start with Hardening_Standard_Register - define your applicable standards")
    logger.info("3. Complete Asset_Type_Hardening_Matrix - map standards to asset types")
    logger.info("4. Begin asset assessments in Asset_Hardening_Assessment")
    logger.info("5. Document control-level detail in Control_Compliance_Detail")
    logger.info("6. Track gaps in Remediation_Tracking")
    logger.info("7. Monitor overall posture via Compliance_Dashboard")
    logger.info("")
    logger.info("IMPORTANT REMINDERS:")
    logger.info("\u2022 This workbook is a TEMPLATE - customize for your organization")
    logger.info("\u2022 Define hardening standards appropriate for your context")
    logger.info("\u2022 Collect evidence systematically - every 'Implemented' control needs evidence")
    logger.info("\u2022 Use exceptions sparingly (<5% of controls)")
    logger.info("\u2022 Prioritize remediation based on risk (see Gap_Prioritization)")
    logger.info("\u2022 Integrate with A.8.9.3 (Monitoring) for continuous compliance")
    logger.info("")
    logger.info("=" * 80)
    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
