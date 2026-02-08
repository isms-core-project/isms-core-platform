#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.9.1 - Baseline Configuration Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Assessment Domain 1 of 4: Baseline Configuration Management

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific configuration management infrastructure, asset
inventory, and baseline documentation requirements.

Key customization areas:
1. Asset types and inventory structure (match your actual infrastructure)
2. Baseline template standards (adapt to your technology stack)
3. Golden image management systems (specific to your deployment tools)
4. Approval workflow roles (align with your organizational structure)
5. Configuration repository locations (based on your CMDB/version control)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
baseline configuration management controls across all asset types against ISO
27001:2022 Control A.8.9 requirements.

**Purpose:**
Enables systematic assessment of baseline definition, documentation, approval,
and maintenance processes to support evidence-based validation of configuration
management effectiveness and audit readiness.

**Assessment Scope:**
- Asset inventory completeness and accuracy
- Baseline configuration documentation for all asset types
- Golden image management and approval
- Baseline approval workflows and authorization
- Configuration repository management
- Deviation tracking and exception management
- Coverage analysis by asset criticality tier
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and baseline standards
2. Asset Inventory - Complete asset inventory by type and tier
3. Baseline Documentation - Baseline definitions and documentation status
4. Golden Images - Golden image inventory and approval tracking
5. Approval Records - Baseline approval workflow evidence
6. Configuration Snapshots - Actual configuration vs. baseline comparison
7. Deviations - Authorized deviations and exceptions
8. Gap Analysis - Coverage gaps and remediation requirements
9. Evidence Register - Audit evidence tracking (100+ entries)
10. Approval & Sign-Off - Three-tier approval workflow

**Key Features:**
- Data validation with asset type and criticality tier dropdowns
- Conditional formatting for baseline coverage status
- Automated gap identification for missing baselines
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with asset inventory systems

**Integration:**
This assessment feeds into the A.8.9.5 Compliance Dashboard, which
consolidates data from all four configuration management assessment domains
for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_1_baseline_configuration.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_1_baseline_configuration.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_1_baseline_configuration.py --date 20250127

Output:
    File: ISMS_IMP_A_8_9_1_Baseline_Configuration_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize asset type taxonomy to match your environment
    2. Import asset inventory from CMDB or network discovery
    3. Document baseline configurations for each asset type
    4. Validate golden image inventory and approval status
    5. Track baseline approval workflow completion
    6. Compare configuration snapshots against baselines
    7. Document authorized deviations with business justification
    8. Conduct gap analysis for missing baselines
    9. Define remediation actions with timelines
    10. Collect and link audit evidence
    11. Obtain three-tier stakeholder approvals
    12. Feed results into A.8.9.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Assessment Domain:    1 of 4 (Baseline Configuration Management)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-POL-A.8.9, Section 2.2: Baseline Configuration Management
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9, Part 1: Technical Standards Reference
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.2: Change Control Assessment (Domain 2)
    - ISMS-IMP-A.8.9.3: Configuration Monitoring Assessment (Domain 3)
    - ISMS-IMP-A.8.9.4: Security Hardening Assessment (Domain 4)
    - ISMS-IMP-A.8.9.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.9.1 specification
    - Supports comprehensive baseline configuration evaluation
    - Integrated with A.8.9.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Baseline Standards:**
Baseline configuration standards evolve with technology changes. Review vendor
security guides, CIS Benchmarks, and DISA STIGs quarterly and update baseline
templates accordingly. Deprecated configurations must be identified and updated.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of baselines, approvals, and coverage metrics.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Complete asset inventory with network topology
- Baseline documentation with security settings
- Configuration snapshots with actual system settings
- Vulnerability information and security gaps

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check asset inventory completeness and baseline coverage
- Semi-annually: Update baseline templates for technology changes
- Annually: Complete reassessment of all asset types
- Ad-hoc: When new asset types deployed or infrastructure changes

**Quality Assurance:**
Have configuration management SMEs and security architects validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Ensure baseline configuration practices align with applicable requirements:
- ISO 27001:2022: Control A.8.9 compliance
- PCI DSS v4.0.1: System configuration documentation requirements
- Sector-specific: Regulatory configuration baseline standards
- Internal: Organizational baseline and hardening policies

Customize assessment criteria to include regulatory-specific requirements.

================================================================================

"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# ============================================================================
# CONFIGURATION SECTION - CUSTOMIZE FOR YOUR ENVIRONMENT
# ============================================================================

# File output configuration
FILENAME = f"ISMS-IMP-A.8.9.1_Baseline_Configuration_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"

# Workbook metadata
WORKBOOK_TITLE = "Baseline Configuration Assessment"
WORKBOOK_VERSION = "1.0"
DOCUMENT_ID = "ISMS-IMP-A.8.9.1"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Baseline_Configuration_{GENERATED_TIMESTAMP}.xlsx"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.8.9: Configuration Management"

# CUSTOMIZE: Dropdown values for your organization
ASSET_CRITICALITY = ["🔴 Critical", "🟡 High", "🟢 Medium", "⭕ Low"]
BASELINE_STATUS = ["\u2705 Defined", "⏳ In Progress", "🔴 Not Started", "➖ N/A"]
APPROVAL_STATUS_BASELINE = ["\u2705 Approved", "🔍 Pending Review", "\u274C Rejected", "📝 Draft"]
APPROVAL_STATUS_TRACKING = ["📝 Pending Submission", "📤 Submitted", "🔍 Under Review", "\u2705 Approved", "\u274C Rejected", "\u1F4CB Revisions Requested"]
APPROVAL_METHOD = ["Change Advisory Board", "Email Approval", "Manager Sign-Off", "Governance Committee", "Automated Process"]
RISK_ASSESSMENT = ["⭕ Low", "🟢 Medium", "🟡 High"]
RISK_ASSESSMENT_DEVIATION = ["⭕ Low", "🟢 Medium", "🟡 High", "🔴 Critical"]

DOC_COMPLETENESS = ["Comprehensive", "Adequate", "Insufficient", "Missing"]
DOC_ACCURACY = ["Verified Accurate", "Mostly Accurate", "Contains Errors", "Not Verified"]
DOC_MAINTAINABILITY = ["Easy to Update", "Moderate Effort", "Difficult", "Not Maintainable"]
DOC_ACCESSIBILITY = ["Highly Accessible", "Accessible", "Limited Access", "Not Accessible"]
REMEDIATION_PRIORITY = ["🔴 Critical", "🟡 High", "🟢 Medium", "⭕ Low"]

CHANGE_TYPE = ["Initial Release", "Minor Update", "Major Update", "Security Patch", "Emergency Change", "Rollback"]
DEVIATION_TYPE = ["Configuration Exception", "Exclusion from Baseline", "Temporary Deviation", "Permanent Exception"]
REVIEW_FREQUENCY = ["Monthly", "Quarterly", "Semi-Annual", "Annual"]
DEVIATION_STATUS = ["\u2705 Active", "🔍 Under Review", "⏰ Expired", "\u274C Revoked", "➖ No Longer Needed"]

EVIDENCE_TYPE = ["Screenshot", "Configuration File", "Scan Report", "Approval Record", "Meeting Minutes", 
                 "Email", "Document", "System Export", "Video Recording", "Audit Log", "Other"]
EVIDENCE_CLASSIFICATION = ["Public", "Internal", "Confidential", "Restricted"]
RETENTION_PERIOD = ["1 Year", "3 Years", "5 Years", "7 Years", "Indefinite"]
VERIFICATION_STATUS = ["\u2705 Verified", "🔍 Needs Verification", "\u274C Missing", "⏰ Outdated"]

APPROVAL_DECISION = ["Approved", "Approved with Conditions", "Not Approved - Revisions Required"]

# CUSTOMIZE: 43-type asset taxonomy
# Organized by category: Infrastructure, Endpoint, Network Services, Applications, Cloud, IoT/OT
ASSET_TYPES = {
    "Infrastructure": [
        "Physical Server",
        "Virtual Machine (VM)",
        "Hypervisor",
        "Network Switch",
        "Network Router",
        "Firewall",
        "Load Balancer",
        "Storage Array (SAN/NAS)",
        "Backup Appliance",
        "UPS/Power Distribution",
        "Environmental Control (HVAC)",
        "Physical Security System"
    ],
    "Endpoint": [
        "Desktop Computer",
        "Laptop Computer",
        "Mobile Phone",
        "Tablet",
        "Thin Client",
        "Kiosk/Point-of-Sale"
    ],
    "Network Services": [
        "DNS Server",
        "DHCP Server",
        "NTP Server",
        "Proxy Server",
        "VPN Gateway",
        "Wireless Access Point",
        "Network Management System"
    ],
    "Applications": [
        "Enterprise Application",
        "Web Application",
        "Database System",
        "Middleware",
        "API/Web Service",
        "Custom Developed Application",
        "Commercial Off-the-Shelf (COTS)",
        "Open Source Software"
    ],
    "Cloud": [
        "IaaS Resource (VM, Storage)",
        "PaaS Service",
        "SaaS Application",
        "Cloud-Native Application",
        "Serverless Function"
    ],
    "IoT/OT": [
        "IoT Device",
        "Industrial Control System (ICS)",
        "SCADA System",
        "Building Management System (BMS)",
        "Medical Device"
    ]
}

# CUSTOMIZE: Color scheme (if different from standard)
COLORS = {
    'header_main': '003366',          # Dark Blue
    'header_sub': '4472C4',           # Blue
    'column_header': 'D9D9D9',        # Light Gray
    'input_cell': 'FFFFFF',           # White (input areas)
    'protected_cell': 'F2F2F2',       # Very Light Gray (protected/calculated)
    'compliant': 'C6EFCE',            # Green
    'partial': 'FFEB9C',              # Yellow
    'non_compliant': 'FFC7CE',        # Red
    'excluded': 'D9D9D9',             # Gray
    'critical': 'C00000',             # Dark Red
    'info_bg': 'E7E6E6'               # Light Blue-Gray (informational)
}

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

def create_styles():
    """
    Creates and returns a dictionary of reusable styles.
    
    Returns:
        dict: Style definitions for different cell types
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    styles = {
        'header_main': {
            'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_main'], 
                              end_color=COLORS['header_main'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'header_sub': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'column_header': {
            'font': Font(name='Calibri', size=11, bold=True),
            'fill': PatternFill(start_color=COLORS['column_header'], 
                              end_color=COLORS['column_header'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'data_cell': {
            'font': Font(name='Calibri', size=11),
            'fill': PatternFill(start_color=COLORS['input_cell'], 
                              end_color=COLORS['input_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False),
            'border': thin_border
        },
        'protected_cell': {
            'font': Font(name='Calibri', size=11, italic=True),
            'fill': PatternFill(start_color=COLORS['protected_cell'], 
                              end_color=COLORS['protected_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top'),
            'border': thin_border
        },
        'info_text': {
            'font': Font(name='Calibri', size=11),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True)
        },
        'section_header': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': thin_border
        }
    }
    return styles

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def apply_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.
    
    Args:
        cell: Cell object to style
        style_dict: Dictionary containing style properties
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def set_column_widths(ws, widths):
    """
    Set column widths for a worksheet.
    
    Args:
        ws: Worksheet object
        widths: Dict mapping column letters to widths
    """
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_data_validation(values, allow_blank=True):
    """
    Create a data validation object for dropdowns.
    
    Args:
        values: List of allowed values
        allow_blank: Whether to allow blank cells
        
    Returns:
        DataValidation object
    """
    formula = f'"{",".join(values)}"'
    
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the dropdown'
    dv.promptTitle = 'Selection Required'
    return dv

def protect_formula_cells(ws, start_row, end_row, formula_columns):
    """
    Protect cells containing formulas while leaving input cells unlocked.
    
    Args:
        ws: Worksheet object
        start_row: First row to protect
        end_row: Last row to protect
        formula_columns: List of column letters containing formulas
    """
    for row in range(start_row, end_row + 1):
        for col in formula_columns:
            cell = ws[f'{col}{row}']
            cell.protection = Protection(locked=True)

# ============================================================================
# LOOKUP TABLE CREATION (HIDDEN SHEET)
# ============================================================================

def create_lookup_tables(wb, styles):
    """
    Create hidden sheet with lookup tables for dropdowns (43-type asset taxonomy).
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Lookup_Tables")
    
    # Create asset type taxonomy with category mapping
    ws['A1'] = "Asset Type"
    ws['B1'] = "Category"
    apply_style(ws['A1'], styles['column_header'])
    apply_style(ws['B1'], styles['column_header'])
    
    row = 2
    for category, types in ASSET_TYPES.items():
        for asset_type in types:
            ws[f'A{row}'] = asset_type
            ws[f'B{row}'] = category
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    # Hide this sheet (users don't need to see it)
    ws.sheet_state = 'hidden'
    
    return ws

# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================

def create_instructions_sheet(wb, styles):
    """
    Create the Instructions and Legend sheet.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Instructions", 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 100
    
    # Title
    ws.merge_cells('A1:A2')
    ws['A1'] = f"{DOCUMENT_ID}  -  Baseline Configuration Assessment\n{CONTROL_REF}"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORS['header_main'],
                                end_color=COLORS['header_main'],
                                fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40
    
    # Document metadata
    ws['A3'] = "Document ID:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = DOCUMENT_ID
    
    ws['A4'] = "Assessment:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Baseline Configuration Assessment"
    
    ws['A5'] = "Version:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "1.0"
    
    ws['A6'] = "Generated:"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime("%d.%m.%Y %H:%M")
    
    ws.column_dimensions['B'].width = 40
    
    # CUSTOMIZE: Update instructions for your organization's specific needs
    instructions = [
        "",
        "ASSESSMENT OVERVIEW",
        "━" * 80,
        "",
        "Purpose:",
        "This workbook assesses the establishment, documentation, approval, and maintenance of configuration "
        "baselines across [Organization]'s information assets. This assessment provides evidence of compliance "
        "with ISO 27001:2022 Control A.8.9 baseline configuration requirements.",
        "",
        "Assessment Scope:",
        "\u2022 All information assets requiring baseline configuration management",
        "\u2022 43 asset types across 6 major categories (Infrastructure, Endpoint, Network, Applications, Cloud, IoT/OT)",
        "\u2022 Baseline documentation quality, approval status, and version control",
        "\u2022 Configuration deviations and exceptions",
        "",
        "WHO SHOULD COMPLETE THIS ASSESSMENT",
        "━" * 80,
        "",
        "Preparer (System Administrators, Asset Owners):",
        "\u2022 Complete Asset_Inventory sheet for all assigned assets",
        "\u2022 Document baseline status and references",
        "\u2022 Collect supporting evidence",
        "\u2022 Complete Evidence_Register",
        "\u2022 Timeline: 2-3 weeks",
        "",
        "Reviewer (Configuration Manager, Team Leads):",
        "\u2022 Verify completeness of Asset_Inventory",
        "\u2022 Complete Documentation_Assessment (quality scoring)",
        "\u2022 Review Approval_Tracking and Deviation_Register",
        "\u2022 Identify remediation priorities",
        "\u2022 Timeline: 1 week",
        "",
        "Approver (IT Manager, CISO):",
        "\u2022 Review Metrics_Summary for overall compliance",
        "\u2022 Assess remediation plans for identified gaps",
        "\u2022 Approve assessment or request revisions",
        "\u2022 Timeline: 3-5 business days",
        "",
        "COMPLETION INSTRUCTIONS",
        "━" * 80,
        "",
        "Step 1 - Asset Inventory (Days 1-5):",
        "1. Export asset list from CMDB if available, or compile manually",
        "2. Open Asset_Inventory sheet",
        "3. For each asset, complete columns A-C (ID, Name, Type), E-G (Criticality, Owner, Location)",
        "4. Determine Baseline Status: Defined, In Progress, Not Started, or N/A",
        "5. If baseline exists, provide reference in column I",
        "6. Update Last Reviewed Date if applicable",
        "",
        "Step 2 - Baseline Documentation (Days 6-10):",
        "1. Open Baseline_Repository sheet",
        "2. Document each distinct configuration baseline",
        "3. Provide baseline version, approval status, documentation location",
        "4. Ensure Baseline IDs match references in Asset_Inventory",
        "",
        "Step 3 - Tracking and Quality (Days 11-14):",
        "1. Complete Approval_Tracking for baselines in approval workflow",
        "2. Configuration Manager completes Documentation_Assessment",
        "3. Update Version_Control with baseline version history",
        "4. Document deviations in Deviation_Register",
        "",
        "Step 4 - Evidence Collection (Days 15-18):",
        "1. Compile all supporting evidence (screenshots, config files, approvals)",
        "2. Document each piece of evidence in Evidence_Register",
        "3. Ensure evidence is accessible for audit verification",
        "",
        "Step 5 - Review and Approval (Days 19-21):",
        "1. Review Metrics_Summary - identify gaps",
        "2. Preparer completes Approval_Sign_Off (Section B)",
        "3. Reviewer verifies and completes Section C",
        "4. Approver reviews and completes Section D",
        "",
        "WORKBOOK STRUCTURE",
        "━" * 80,
        "",
        "This workbook contains 11 sheets:",
        "",
        "1.  Instructions - This sheet (usage guidance)",
        "2.  Asset_Inventory - All assets requiring baseline management (100 rows)",
        "3.  Baseline_Repository - Documented configuration baselines (50 rows)",
        "4.  Baseline_Coverage_Matrix - Coverage analysis by asset type (43 rows, auto-calculated)",
        "5.  Approval_Tracking - Baseline approval workflow status (50 rows)",
        "6.  Documentation_Assessment - Quality evaluation of baseline docs (30 rows)",
        "7.  Version_Control - Baseline version history (50 rows)",
        "8.  Deviation_Register - Authorized deviations from baselines (50 rows)",
        "9.  Metrics_Summary - Auto-calculated compliance metrics (dashboard)",
        "10. Evidence_Register - Supporting evidence documentation (100 rows)",
        "11. Approval_Sign_Off - Three-tier approval signatures",
        "",
        "LEGEND - STATUS VALUES AND COLOR CODING",
        "━" * 80,
        "",
        "Baseline Status:",
        "  🟢 Defined (Green) - Baseline is defined, documented, and approved",
        "  🟡 In Progress (Yellow) - Baseline definition is underway",
        "  🔴 Not Started (Red) - Baseline has not been initiated",
        "  ⚪ N/A (Gray) - Baseline not applicable (with justification)",
        "",
        "Approval Status:",
        "  🟢 Approved (Green) - Baseline is formally approved",
        "  🟡 Under Review (Yellow) - Currently in approval process",
        "  🔴 Rejected (Red) - Baseline was rejected, needs rework",
        "  🟠 Revisions Requested (Orange) - Changes required before approval",
        "",
        "Documentation Quality:",
        "  Excellent (≥90 score) - Comprehensive, accurate, maintainable, accessible",
        "  Good (75-89 score) - Adequate quality with minor gaps",
        "  Fair (50-74 score) - Acceptable but improvement needed",
        "  Poor (<50 score) - Significant quality issues, requires remediation",
        "",
        "Compliance Status:",
        "  🟢 Compliant - Meets or exceeds target coverage",
        "  🟡 Partial - Below target but within acceptable range",
        "  🔴 Non-Compliant - Significantly below target, action required",
        "",
        "IMPORTANT NOTES",
        "━" * 80,
        "",
        "\u2022 Protected cells (gray background) contain formulas - do not edit",
        "\u2022 Use dropdowns for standardized entries - do not type free text where dropdown exists",
        "\u2022 All date fields use DD.MM.YYYY format",
        "\u2022 Baseline_Coverage_Matrix and Metrics_Summary update automatically",
        "\u2022 Evidence_Register is critical for audit verification",
        "\u2022 Complete Approval_Sign_Off only when entire assessment is finalized",
        "",
        "COMPLIANCE TARGETS",
        "━" * 80,
        "",
        "Critical Assets: ≥95% baseline coverage",
        "High Assets: ≥90% baseline coverage",
        "Medium Assets: ≥80% baseline coverage",
        "Low Assets: ≥60% baseline coverage",
        "Overall: ≥85% baseline coverage",
        "",
        "Documentation Quality: ≥75% of baselines should have 'Good' or 'Excellent' rating",
        "Approval SLA: <14 days from submission to approval decision",
        "High/Critical Risk Deviations: <5 active",
        "",
        "SUPPORT AND QUESTIONS",
        "━" * 80,
        "",
        "Configuration Manager: [contact information]",
        "ISMS Team: [contact information]",
        "",
        "For technical issues with this workbook, contact: [IT support]",
        "For questions about Control A.8.9 requirements, reference: ISMS-POL-A.8.9",
        "",
        "━" * 80,
        f"Generated: {datetime.now().strftime('%d.%m.%Y')} | Version: {WORKBOOK_VERSION} | Document ID: {DOCUMENT_ID}",
    ]
    
    row = 8
    for line in instructions:
        ws[f'A{row}'] = line
        if line.startswith("━"):
            ws[f'A{row}'].font = Font(name='Calibri', size=11, color='666666')
        elif line.isupper() and len(line) > 5 and not line.startswith("  "):
            apply_style(ws[f'A{row}'], styles['section_header'])
            ws.row_dimensions[row].height = 20
        else:
            apply_style(ws[f'A{row}'], styles['info_text'])
        row += 1
    
    return ws

def create_asset_inventory_sheet(wb, styles):
    """
    Create Sheet 2: Asset_Inventory
    
    Purpose: Comprehensive list of all information assets requiring baseline management.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Asset_Inventory")
    
    # Set column widths
    widths = {
        'A': 15,  # Asset ID
        'B': 30,  # Asset Name
        'C': 35,  # Asset Type
        'D': 20,  # Asset Category
        'E': 12,  # Criticality
        'F': 20,  # Owner
        'G': 25,  # Location
        'H': 15,  # Baseline Status
        'I': 25,  # Baseline Reference
        'J': 15,  # Last Reviewed Date
        'K': 15,  # Next Review Due
        'L': 15,  # Compliance Status
        'M': 40   # Notes
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:M1')
    ws['A1'] = "Asset Inventory - Baseline Configuration Assessment"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Asset ID',
        'B': 'Asset Name',
        'C': 'Asset Type',
        'D': 'Asset Category',
        'E': 'Criticality',
        'F': 'Owner',
        'G': 'Location',
        'H': 'Baseline Status',
        'I': 'Baseline Reference',
        'J': 'Last Reviewed Date',
        'K': 'Next Review Due',
        'L': 'Compliance Status',
        'M': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Create 100 data rows
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            # Formula columns get protected style
            if col in ['D', 'K', 'L']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add formulas to protected columns
    for row in range(3, 3 + num_rows):
        # Column D: Asset Category (lookup from Lookup_Tables)
        ws[f'D{row}'] = f'=IF(C{row}="","",IFERROR(VLOOKUP(C{row},Lookup_Tables!$A$2:$B$44,2,FALSE),"Unknown"))'
        
        # Column K: Next Review Due (based on criticality)
        ws[f'K{row}'] = f'=IF(J{row}="","",IF(E{row}="Critical",J{row}+90,IF(E{row}="High",J{row}+90,IF(E{row}="Medium",J{row}+180,J{row}+180))))'
        
        # Column L: Compliance Status
        ws[f'L{row}'] = f'=IF(H{row}="\u2705 Defined","\u2705 Compliant",IF(H{row}="➖ N/A","➖ Excluded","\u274C Non-Compliant"))'
    
    # Data validations
    # Asset Type (dropdown from Lookup_Tables)
    asset_type_dv = DataValidation(type="list", formula1="=Lookup_Tables!$A$2:$A$44", allow_blank=False)
    asset_type_dv.error = 'Please select a valid asset type'
    asset_type_dv.errorTitle = 'Invalid Asset Type'
    ws.add_data_validation(asset_type_dv)
    asset_type_dv.add(f'C3:C{2+num_rows}')
    
    # Criticality
    criticality_dv = create_data_validation(ASSET_CRITICALITY, allow_blank=False)
    ws.add_data_validation(criticality_dv)
    criticality_dv.add(f'E3:E{2+num_rows}')
    
    # Baseline Status
    baseline_status_dv = create_data_validation(BASELINE_STATUS, allow_blank=False)
    ws.add_data_validation(baseline_status_dv)
    baseline_status_dv.add(f'H3:H{2+num_rows}')
    
    # Conditional formatting - Baseline Status column
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Defined"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"In Progress"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Started"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"N/A"'], 
                   fill=PatternFill(start_color=COLORS['excluded'], end_color=COLORS['excluded'], fill_type='solid')))
    
    # Conditional formatting - Compliance Status column
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='containsText', formula=['"Compliant"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='containsText', formula=['"Non-Compliant"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    # Protect formula cells
    protect_formula_cells(ws, 3, 2 + num_rows, ['D', 'K', 'L'])
    
    return ws

def create_baseline_repository_sheet(wb, styles):
    """
    Create Sheet 3: Baseline_Repository
    
    Purpose: Catalog of all configuration baselines maintained by [Organization].
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Baseline_Repository")
    
    # Set column widths
    widths = {
        'A': 20,  # Baseline ID
        'B': 35,  # Baseline Name
        'C': 12,  # Version
        'D': 30,  # Applicable Asset Types
        'E': 50,  # Description
        'F': 18,  # Approval Status
        'G': 20,  # Approved By
        'H': 15,  # Approval Date
        'I': 15,  # Last Updated
        'J': 40,  # Documentation Location
        'K': 12,  # Config Elements Count
        'L': 12,  # Applied to Assets Count
        'M': 40   # Notes
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:M1')
    ws['A1'] = "Baseline Repository"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Baseline ID',
        'B': 'Baseline Name',
        'C': 'Baseline Version',
        'D': 'Applicable Asset Types',
        'E': 'Description',
        'F': 'Approval Status',
        'G': 'Approved By',
        'H': 'Approval Date',
        'I': 'Last Updated',
        'J': 'Documentation Location',
        'K': 'Config Elements Count',
        'L': 'Applied to Assets Count',
        'M': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Create 50 data rows
    num_rows = 50
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            # Column L is optional formula column
            if col == 'L':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add optional formula to column L (Applied to Assets Count)
    # This counts how many assets reference this baseline ID
    for row in range(3, 3 + num_rows):
        ws[f'L{row}'] = f'=COUNTIF(Asset_Inventory!$I$3:$I$102,A{row})'
    
    # Data validations
    approval_status_dv = create_data_validation(APPROVAL_STATUS_BASELINE, allow_blank=False)
    ws.add_data_validation(approval_status_dv)
    approval_status_dv.add(f'F3:F{2+num_rows}')
    
    # Conditional formatting - Approval Status
    ws.conditional_formatting.add(f'F3:F{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Approved"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'F3:F{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Pending Review"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'F3:F{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Rejected"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    # Protect formula cells
    protect_formula_cells(ws, 3, 2 + num_rows, ['L'])
    
    return ws

def create_baseline_coverage_matrix_sheet(wb, styles):
    """
    Create Sheet 4: Baseline_Coverage_Matrix
    
    Purpose: Statistical analysis of baseline coverage by asset type (formula-driven).
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Baseline_Coverage_Matrix")
    
    # Set column widths
    widths = {
        'A': 40,  # Asset Type
        'B': 20,  # Asset Category
        'C': 12,  # Total Assets
        'D': 15,  # Assets with Baselines
        'E': 15,  # Assets In Progress
        'F': 15,  # Assets Not Started
        'G': 12,  # Assets N/A
        'H': 12,  # Coverage %
        'I': 12,  # Critical Assets Count
        'J': 15,  # Critical Coverage %
        'K': 15,  # Status
        'L': 40   # Gap Analysis
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:L1')
    ws['A1'] = "Baseline Coverage Matrix by Asset Type"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Asset Type',
        'B': 'Asset Category',
        'C': 'Total Assets',
        'D': 'Assets with Baselines',
        'E': 'Assets In Progress',
        'F': 'Assets Not Started',
        'G': 'Assets N/A',
        'H': 'Coverage %',
        'I': 'Critical Assets Count',
        'J': 'Critical Coverage %',
        'K': 'Status',
        'L': 'Gap Analysis'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Pre-fill 43 asset types from taxonomy
    row = 3
    for category, types in ASSET_TYPES.items():
        for asset_type in types:
            ws[f'A{row}'] = asset_type
            ws[f'B{row}'] = category
            
            # Add formulas for all metric columns
            # Column C: Total Assets
            ws[f'C{row}'] = f'=COUNTIF(Asset_Inventory!$C$3:$C$102,A{row})'
            
            # Column D: Assets with Baselines
            ws[f'D{row}'] = f'=COUNTIFS(Asset_Inventory!$C$3:$C$102,A{row},Asset_Inventory!$H$3:$H$102,"\u2705 Defined")'
            
            # Column E: Assets In Progress
            ws[f'E{row}'] = f'=COUNTIFS(Asset_Inventory!$C$3:$C$102,A{row},Asset_Inventory!$H$3:$H$102,"⏳ In Progress")'
            
            # Column F: Assets Not Started
            ws[f'F{row}'] = f'=COUNTIFS(Asset_Inventory!$C$3:$C$102,A{row},Asset_Inventory!$H$3:$H$102,"🔴 Not Started")'
            
            # Column G: Assets N/A
            ws[f'G{row}'] = f'=COUNTIFS(Asset_Inventory!$C$3:$C$102,A{row},Asset_Inventory!$H$3:$H$102,"➖ N/A")'
            
            # Column H: Coverage %
            ws[f'H{row}'] = f'=IF((C{row}-G{row})=0,0,D{row}/(C{row}-G{row})*100)'
            ws[f'H{row}'].number_format = '0.0'
            
            # Column I: Critical Assets Count
            ws[f'I{row}'] = f'=COUNTIFS(Asset_Inventory!$C$3:$C$102,A{row},Asset_Inventory!$E$3:$E$102,"🔴 Critical")'
            
            # Column J: Critical Coverage %
            ws[f'J{row}'] = f'=IF(I{row}=0,0,COUNTIFS(Asset_Inventory!$C$3:$C$102,A{row},Asset_Inventory!$E$3:$E$102,"🔴 Critical",Asset_Inventory!$H$3:$H$102,"\u2705 Defined")/I{row}*100)'
            ws[f'J{row}'].number_format = '0.0'
            
            # Column K: Status
            ws[f'K{row}'] = f'=IF(H{row}>=90,"\u2705 Compliant",IF(H{row}>=60,"\u26A0\uFE0F Partial","\u274C Non-Compliant"))'
            
            # Column L: Gap Analysis
            ws[f'L{row}'] = f'=IF(ISNUMBER(SEARCH("Compliant",K{row})),"\u2705 Coverage target met",IF(F{row}>0,F{row}&" assets need baselines started","⏳ Review in-progress baselines"))'
            
            # Apply protected style to all cells (formula-driven sheet)
            for col in headers.keys():
                apply_style(ws[f'{col}{row}'], styles['protected_cell'])
            
            row += 1
    
    # Add summary row at bottom
    summary_row = row
    ws[f'A{summary_row}'] = "TOTAL / OVERALL"
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    ws[f'C{summary_row}'] = f'=SUM(C3:C{row-1})'
    ws[f'D{summary_row}'] = f'=SUM(D3:D{row-1})'
    ws[f'E{summary_row}'] = f'=SUM(E3:E{row-1})'
    ws[f'F{summary_row}'] = f'=SUM(F3:F{row-1})'
    ws[f'G{summary_row}'] = f'=SUM(G3:G{row-1})'
    ws[f'H{summary_row}'] = f'=IF((C{summary_row}-G{summary_row})=0,0,D{summary_row}/(C{summary_row}-G{summary_row})*100)'
    ws[f'H{summary_row}'].number_format = '0.0'
    ws[f'I{summary_row}'] = f'=SUM(I3:I{row-1})'
    ws[f'J{summary_row}'] = f'=IF(I{summary_row}=0,0,COUNTIFS(Asset_Inventory!$E$3:$E$102,"🔴 Critical",Asset_Inventory!$H$3:$H$102,"\u2705 Defined")/I{summary_row}*100)'
    ws[f'J{summary_row}'].number_format = '0.0'
    ws[f'K{summary_row}'] = f'=IF(H{summary_row}>=85,"\u2705 Compliant",IF(H{summary_row}>=70,"\u26A0\uFE0F Partial","\u274C Non-Compliant"))'
    
    for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        apply_style(ws[f'{col}{summary_row}'], styles['section_header'])
    
    # Conditional formatting - Coverage %
    ws.conditional_formatting.add(f'H3:H{summary_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['90'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'H3:H{summary_row}',
        CellIsRule(operator='between', formula=['60', '89.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'H3:H{summary_row}',
        CellIsRule(operator='lessThan', formula=['60'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    # Conditional formatting - Critical Coverage %
    ws.conditional_formatting.add(f'J3:J{summary_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J3:J{summary_row}',
        CellIsRule(operator='between', formula=['80', '94.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'J3:J{summary_row}',
        CellIsRule(operator='lessThan', formula=['80'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    # Conditional formatting - Status
    ws.conditional_formatting.add(f'K3:K{summary_row}',
        CellIsRule(operator='containsText', formula=['"Compliant"'], 
                   font=Font(bold=True, color='006100')))
    ws.conditional_formatting.add(f'K3:K{summary_row}',
        CellIsRule(operator='containsText', formula=['"Partial"'], 
                   font=Font(bold=True, color='9C6500')))
    ws.conditional_formatting.add(f'K3:K{summary_row}',
        CellIsRule(operator='containsText', formula=['"Non-Compliant"'], 
                   font=Font(bold=True, color='9C0006')))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    # Protect all cells (completely formula-driven)
    for row_num in range(3, summary_row + 1):
        for col in headers.keys():
            ws[f'{col}{row_num}'].protection = Protection(locked=True)
    
    return ws

def create_approval_tracking_sheet(wb, styles):
    """
    Create Sheet 5: Approval_Tracking
    
    Purpose: Track approval workflow status for each configuration baseline.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Approval_Tracking")
    
    # Set column widths
    widths = {
        'A': 20,  # Baseline ID
        'B': 30,  # Baseline Name
        'C': 15,  # Submission Date
        'D': 20,  # Approver Name
        'E': 20,  # Approval Status
        'F': 15,  # Approval Date
        'G': 25,  # Approval Method
        'H': 30,  # Approval Reference
        'I': 40,  # Business Justification
        'J': 15,  # Risk Assessment
        'K': 12,  # Days Pending
        'L': 15,  # SLA Status
        'M': 30,  # Next Action
        'N': 40   # Notes
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:N1')
    ws['A1'] = "Baseline Approval Tracking"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Baseline ID',
        'B': 'Baseline Name',
        'C': 'Submission Date',
        'D': 'Approver Name',
        'E': 'Approval Status',
        'F': 'Approval Date',
        'G': 'Approval Method',
        'H': 'Approval Reference',
        'I': 'Business Justification',
        'J': 'Risk Assessment',
        'K': 'Days Pending',
        'L': 'SLA Status',
        'M': 'Next Action',
        'N': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Create 50 data rows
    num_rows = 50
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            # Columns K and L are formula columns
            if col in ['K', 'L']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add formulas to protected columns
    for row in range(3, 3 + num_rows):
        # Column K: Days Pending
        ws[f'K{row}'] = f'=IF(C{row}="","",IF(F{row}="",TODAY()-C{row},F{row}-C{row}))'
        
        # Column L: SLA Status (14 day SLA)
        ws[f'L{row}'] = f'=IF(K{row}="","",IF(K{row}<=14,"Within SLA",IF(K{row}<=21,"Approaching SLA","SLA Breach")))'
    
    # Data validations
    approval_status_tracking_dv = create_data_validation(APPROVAL_STATUS_TRACKING, allow_blank=False)
    ws.add_data_validation(approval_status_tracking_dv)
    approval_status_tracking_dv.add(f'E3:E{2+num_rows}')
    
    approval_method_dv = create_data_validation(APPROVAL_METHOD, allow_blank=True)
    ws.add_data_validation(approval_method_dv)
    approval_method_dv.add(f'G3:G{2+num_rows}')
    
    risk_assessment_dv = create_data_validation(RISK_ASSESSMENT, allow_blank=True)
    ws.add_data_validation(risk_assessment_dv)
    risk_assessment_dv.add(f'J3:J{2+num_rows}')
    
    # Conditional formatting - Approval Status
    ws.conditional_formatting.add(f'E3:E{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Approved"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'E3:E{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Under Review"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'E3:E{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Rejected"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    # Conditional formatting - SLA Status
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Within SLA"'], 
                   font=Font(color='006100')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Approaching SLA"'], 
                   font=Font(bold=True, color='9C6500')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"SLA Breach"'], 
                   font=Font(bold=True, color='9C0006')))
    
    # Conditional formatting - Days Pending (>21 days)
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='greaterThan', formula=['21'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    # Protect formula cells
    protect_formula_cells(ws, 3, 2 + num_rows, ['K', 'L'])
    
    return ws

def create_documentation_assessment_sheet(wb, styles):
    """
    Create Sheet 6: Documentation_Assessment
    
    Purpose: Evaluate quality of baseline documentation against defined criteria.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Documentation_Assessment")
    
    # Set column widths
    widths = {
        'A': 20,  # Baseline ID
        'B': 30,  # Baseline Name
        'C': 18,  # Documentation Completeness
        'D': 12,  # Completeness Score
        'E': 18,  # Documentation Accuracy
        'F': 12,  # Accuracy Score
        'G': 18,  # Maintainability
        'H': 12,  # Maintainability Score
        'I': 18,  # Accessibility
        'J': 12,  # Accessibility Score
        'K': 15,  # Overall Quality Score
        'L': 15,  # Quality Rating
        'M': 40,  # Gaps Identified
        'N': 15,  # Remediation Priority
        'O': 15   # Target Completion Date
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:O1')
    ws['A1'] = "Baseline Documentation Quality Assessment"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Baseline ID',
        'B': 'Baseline Name',
        'C': 'Documentation Completeness',
        'D': 'Completeness Score',
        'E': 'Documentation Accuracy',
        'F': 'Accuracy Score',
        'G': 'Maintainability',
        'H': 'Maintainability Score',
        'I': 'Accessibility',
        'J': 'Accessibility Score',
        'K': 'Overall Quality Score',
        'L': 'Quality Rating',
        'M': 'Gaps Identified',
        'N': 'Remediation Priority',
        'O': 'Target Completion Date'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Create 30 data rows
    num_rows = 30
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            # Score columns and rating columns are formulas
            if col in ['D', 'F', 'H', 'J', 'K', 'L']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add formulas to score columns
    for row in range(3, 3 + num_rows):
        # Column D: Completeness Score
        ws[f'D{row}'] = f'=IF(C{row}="Comprehensive",100,IF(C{row}="Adequate",75,IF(C{row}="Insufficient",40,IF(C{row}="Missing",0,""))))'
        
        # Column F: Accuracy Score
        ws[f'F{row}'] = f'=IF(E{row}="Verified Accurate",100,IF(E{row}="Mostly Accurate",75,IF(E{row}="Contains Errors",40,IF(E{row}="Not Verified",0,""))))'
        
        # Column H: Maintainability Score
        ws[f'H{row}'] = f'=IF(G{row}="Easy to Update",100,IF(G{row}="Moderate Effort",75,IF(G{row}="Difficult",40,IF(G{row}="Not Maintainable",0,""))))'
        
        # Column J: Accessibility Score
        ws[f'J{row}'] = f'=IF(I{row}="Highly Accessible",100,IF(I{row}="Accessible",75,IF(I{row}="Limited Access",40,IF(I{row}="Not Accessible",0,""))))'
        
        # Column K: Overall Quality Score (average of four dimensions)
        ws[f'K{row}'] = f'=IF(AND(D{row}="",F{row}="",H{row}="",J{row}=""),"",(D{row}+F{row}+H{row}+J{row})/4)'
        ws[f'K{row}'].number_format = '0.0'
        
        # Column L: Quality Rating
        ws[f'L{row}'] = f'=IF(K{row}="","",IF(K{row}>=90,"Excellent",IF(K{row}>=75,"Good",IF(K{row}>=50,"Fair","Poor"))))'
    
    # Data validations
    completeness_dv = create_data_validation(DOC_COMPLETENESS, allow_blank=False)
    ws.add_data_validation(completeness_dv)
    completeness_dv.add(f'C3:C{2+num_rows}')
    
    accuracy_dv = create_data_validation(DOC_ACCURACY, allow_blank=False)
    ws.add_data_validation(accuracy_dv)
    accuracy_dv.add(f'E3:E{2+num_rows}')
    
    maintainability_dv = create_data_validation(DOC_MAINTAINABILITY, allow_blank=False)
    ws.add_data_validation(maintainability_dv)
    maintainability_dv.add(f'G3:G{2+num_rows}')
    
    accessibility_dv = create_data_validation(DOC_ACCESSIBILITY, allow_blank=False)
    ws.add_data_validation(accessibility_dv)
    accessibility_dv.add(f'I3:I{2+num_rows}')
    
    remediation_priority_dv = create_data_validation(REMEDIATION_PRIORITY, allow_blank=True)
    ws.add_data_validation(remediation_priority_dv)
    remediation_priority_dv.add(f'N3:N{2+num_rows}')
    
    # Conditional formatting - Overall Quality Score
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='greaterThanOrEqual', formula=['90'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='between', formula=['75', '89.9'], 
                   fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='between', formula=['50', '74.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='lessThan', formula=['50'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    # Conditional formatting - Quality Rating
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Excellent"'], 
                   font=Font(bold=True, color='006100')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Good"'], 
                   font=Font(color='006100')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Fair"'], 
                   font=Font(color='9C6500')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Poor"'], 
                   font=Font(bold=True, color='9C0006')))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    # Protect formula cells
    protect_formula_cells(ws, 3, 2 + num_rows, ['D', 'F', 'H', 'J', 'K', 'L'])
    
    return ws

def create_version_control_sheet(wb, styles):
    """
    Create Sheet 7: Version_Control
    
    Purpose: Track version history of configuration baselines over time.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Version_Control")
    
    # Set column widths
    widths = {
        'A': 20,  # Baseline ID
        'B': 30,  # Baseline Name
        'C': 15,  # Version Number
        'D': 15,  # Version Date
        'E': 15,  # Previous Version
        'F': 18,  # Change Type
        'G': 40,  # Change Summary
        'H': 25,  # Change Request Reference
        'I': 20,  # Changed By
        'J': 20,  # Approved By
        'K': 15,  # Superseded Date
        'L': 12,  # Status
        'M': 12,  # Assets Affected Count
        'N': 40   # Documentation Location
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:N1')
    ws['A1'] = "Baseline Version Control History"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Baseline ID',
        'B': 'Baseline Name',
        'C': 'Version Number',
        'D': 'Version Date',
        'E': 'Previous Version',
        'F': 'Change Type',
        'G': 'Change Summary',
        'H': 'Change Request Reference',
        'I': 'Changed By',
        'J': 'Approved By',
        'K': 'Superseded Date',
        'L': 'Status',
        'M': 'Assets Affected Count',
        'N': 'Documentation Location'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Create 50 data rows
    num_rows = 50
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            # Column L is formula column
            if col == 'L':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add formulas to column L (Status)
    for row in range(3, 3 + num_rows):
        ws[f'L{row}'] = f'=IF(K{row}="","Current","Superseded")'
    
    # Data validations
    change_type_dv = create_data_validation(CHANGE_TYPE, allow_blank=False)
    ws.add_data_validation(change_type_dv)
    change_type_dv.add(f'F3:F{2+num_rows}')
    
    # Conditional formatting - Status
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Current"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Superseded"'], 
                   fill=PatternFill(start_color=COLORS['excluded'], end_color=COLORS['excluded'], fill_type='solid')))
    
    # Conditional formatting - Change Type (highlight security and emergency)
    ws.conditional_formatting.add(f'F3:F{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Emergency Change"'], 
                   font=Font(color='9C0006')))
    ws.conditional_formatting.add(f'F3:F{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Security Patch"'], 
                   font=Font(color='C65911')))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    # Protect formula cells
    protect_formula_cells(ws, 3, 2 + num_rows, ['L'])
    
    return ws

def create_deviation_register_sheet(wb, styles):
    """
    Create Sheet 8: Deviation_Register
    
    Purpose: Document authorized deviations from standard configuration baselines.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Deviation_Register")
    
    # Set column widths
    widths = {
        'A': 18,  # Deviation ID
        'B': 15,  # Asset ID
        'C': 25,  # Asset Name
        'D': 20,  # Baseline ID
        'E': 25,  # Deviation Type
        'F': 30,  # Configuration Element
        'G': 25,  # Standard Value
        'H': 25,  # Actual Value
        'I': 40,  # Business Justification
        'J': 12,  # Risk Assessment
        'K': 35,  # Compensating Controls
        'L': 20,  # Approved By
        'M': 15,  # Approval Date
        'N': 15,  # Review Frequency
        'O': 15,  # Next Review Date
        'P': 15,  # Deviation Status
        'Q': 15,  # Expiration Date
        'R': 35   # Notes
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:R1')
    ws['A1'] = "Configuration Baseline Deviation Register"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Deviation ID',
        'B': 'Asset ID',
        'C': 'Asset Name',
        'D': 'Baseline ID',
        'E': 'Deviation Type',
        'F': 'Configuration Element',
        'G': 'Standard Value',
        'H': 'Actual Value',
        'I': 'Business Justification',
        'J': 'Risk Assessment',
        'K': 'Compensating Controls',
        'L': 'Approved By',
        'M': 'Approval Date',
        'N': 'Review Frequency',
        'O': 'Next Review Date',
        'P': 'Deviation Status',
        'Q': 'Expiration Date',
        'R': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Create 50 data rows
    num_rows = 50
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            # Column O is formula column
            if col == 'O':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add formulas to column O (Next Review Date)
    for row in range(3, 3 + num_rows):
        ws[f'O{row}'] = f'=IF(M{row}="","",IF(N{row}="Monthly",M{row}+30,IF(N{row}="Quarterly",M{row}+90,IF(N{row}="Semi-Annual",M{row}+180,IF(N{row}="Annual",M{row}+365,"")))))'
    
    # Data validations
    deviation_type_dv = create_data_validation(DEVIATION_TYPE, allow_blank=False)
    ws.add_data_validation(deviation_type_dv)
    deviation_type_dv.add(f'E3:E{2+num_rows}')
    
    risk_assessment_deviation_dv = create_data_validation(RISK_ASSESSMENT_DEVIATION, allow_blank=False)
    ws.add_data_validation(risk_assessment_deviation_dv)
    risk_assessment_deviation_dv.add(f'J3:J{2+num_rows}')
    
    review_frequency_dv = create_data_validation(REVIEW_FREQUENCY, allow_blank=False)
    ws.add_data_validation(review_frequency_dv)
    review_frequency_dv.add(f'N3:N{2+num_rows}')
    
    deviation_status_dv = create_data_validation(DEVIATION_STATUS, allow_blank=False)
    ws.add_data_validation(deviation_status_dv)
    deviation_status_dv.add(f'P3:P{2+num_rows}')
    
    # Conditional formatting - Risk Assessment
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'], 
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"High"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Medium"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Low"'], 
                   fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
    
    # Conditional formatting - Deviation Status
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Active"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Expired"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Under Review"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    # Protect formula cells
    protect_formula_cells(ws, 3, 2 + num_rows, ['O'])
    
    return ws

def create_metrics_summary_sheet(wb, styles):
    """
    Create Sheet 9: Metrics_Summary
    
    Purpose: Auto-calculate key compliance metrics and executive summary (formula-driven dashboard).
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Metrics_Summary")
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Baseline Configuration Assessment - Metrics Summary"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # SECTION A: Overall Compliance Metrics
    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL COMPLIANCE METRICS"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    # Header row for metrics table
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    # CUSTOMIZE: Metrics formulas reference other sheets
    metrics = [
        ("Total Assets in Scope", "=COUNTA(Asset_Inventory!A3:A102)-COUNTBLANK(Asset_Inventory!A3:A102)", "N/A", ""),
        ("Assets with Defined Baselines", '=COUNTIF(Asset_Inventory!H3:H102,"Defined")', "N/A", ""),
        ("Overall Baseline Coverage %", "=IF(B5=0,0,B6/B5*100)", "≥85%", '=IF(B7>=85,"✓ Compliant",IF(B7>=70,"\u26A0 Partial","✗ Non-Compliant"))'),
        ("Critical Asset Count", '=COUNTIF(Asset_Inventory!E3:E102,"Critical")', "N/A", ""),
        ("Critical Assets with Baselines", '=COUNTIFS(Asset_Inventory!E3:E102,"Critical",Asset_Inventory!H3:H102,"Defined")', "N/A", ""),
        ("Critical Asset Coverage %", "=IF(B8=0,0,B9/B8*100)", "≥95%", '=IF(B10>=95,"✓ Compliant",IF(B10>=90,"\u26A0 Partial","✗ Non-Compliant"))'),
        ("High Asset Count", '=COUNTIF(Asset_Inventory!E3:E102,"High")', "N/A", ""),
        ("High Assets with Baselines", '=COUNTIFS(Asset_Inventory!E3:E102,"High",Asset_Inventory!H3:H102,"Defined")', "N/A", ""),
        ("High Asset Coverage %", "=IF(B11=0,0,B12/B11*100)", "≥90%", '=IF(B13>=90,"✓ Compliant",IF(B13>=85,"\u26A0 Partial","✗ Non-Compliant"))'),
        ("Baselines Pending Approval", '=COUNTIF(Approval_Tracking!E3:E52,"Submitted")+COUNTIF(Approval_Tracking!E3:E52,"Under Review")', "0", '=IF(B14=0,"✓ None Pending","\u26A0 "&B14&" Pending")'),
        ("Baselines with Excellent Documentation", '=COUNTIF(Documentation_Assessment!L3:L32,"Excellent")', "Maximize", '=B15&" baselines"'),
        ("Baselines with Poor Documentation", '=COUNTIF(Documentation_Assessment!L3:L32,"Poor")', "0", '=IF(B16=0,"✓ None","✗ "&B16&" Need Remediation")'),
        ("Active High/Critical Risk Deviations", '=COUNTIFS(Deviation_Register!J3:J52,"High",Deviation_Register!P3:P52,"Active")+COUNTIFS(Deviation_Register!J3:J52,"Critical",Deviation_Register!P3:P52,"Active")', "<5", '=IF(B17<5,"✓ Within Target","✗ "&B17&" High-Risk")'),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = value_formula
        ws[f'B{row}'].number_format = '0.0' if 'Coverage %' in metric_name else '0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION B: Coverage by Asset Category
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "COVERAGE BY ASSET CATEGORY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Category"
    ws[f'B{row}'] = "Total Assets"
    ws[f'C{row}'] = "With Baselines"
    ws[f'D{row}'] = "Coverage %"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    categories = ["Infrastructure", "Endpoint", "Network Services", "Applications", "Cloud", "IoT/OT"]
    row += 1
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        # Total assets in this category
        ws[f'B{row}'] = f'=COUNTIF(Asset_Inventory!$D$3:$D$102,"{category}")'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        # Assets with baselines in this category
        ws[f'C{row}'] = f'=COUNTIFS(Asset_Inventory!$D$3:$D$102,"{category}",Asset_Inventory!$H$3:$H$102,"Defined")'
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        # Coverage %
        ws[f'D{row}'] = f'=IF(B{row}=0,0,C{row}/B{row}*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # Conditional formatting for coverage percentages in Section B
    start_row = row - len(categories)
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['85'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='between', formula=['70', '84.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='lessThan', formula=['70'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    # SECTION C: Approval Process Health
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "APPROVAL PROCESS HEALTH"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    approval_metrics = [
        ("Average Approval Time (Days)", "=IFERROR(AVERAGE(Approval_Tracking!K3:K52),0)", "<14 days", '=IF(B{row}<=14,"✓ Within SLA","✗ Exceeds SLA")'),
        ("Baselines Exceeding SLA (>21 days)", '=COUNTIF(Approval_Tracking!L3:L52,"SLA Breach")', "0", '=IF(B{row}=0,"✓ None","✗ "&B{row}&" Breaches")'),
        ("Total Approved Baselines", '=COUNTIF(Approval_Tracking!E3:E52,"Approved")', "N/A", ""),
        ("Total Rejected Baselines", '=COUNTIF(Approval_Tracking!E3:E52,"Rejected")', "N/A", ""),
        ("Approval Success Rate %", "=IF((B{row_approved}+B{row_rejected})=0,0,B{row_approved}/(B{row_approved}+B{row_rejected})*100)", ">90%", '=IF(B{row}>=90,"✓ Good",IF(B{row}>=80,"\u26A0 Fair","✗ Poor"))'),
    ]
    
    row += 1
    row_approved = row + 2  # Will reference approved count
    row_rejected = row + 3  # Will reference rejected count
    
    for idx, (metric_name, value_formula, target, status_formula) in enumerate(approval_metrics):
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        # Adjust formulas that reference other rows
        if "{row" in value_formula:  # Matches {row}, {row_approved}, {row_rejected}
            value_formula = value_formula.replace("{row_approved}", str(row_approved)).replace("{row_rejected}", str(row_rejected)).replace("{row}", str(row))
        
        ws[f'B{row}'] = value_formula
        if "%" in metric_name or "Days" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        if status_formula:
            status_formula = status_formula.replace("{row_approved}", str(row_approved)).replace("{row_rejected}", str(row_rejected)).replace("{row}", str(row))
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # Add informational note at bottom
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "NOTE: All metrics on this sheet are auto-calculated from data in other sheets. Do not edit."
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Status indicators: ✓ = Compliant, \u26A0 = Partial/Warning, ✗ = Non-Compliant"
    ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='999999')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    return ws

def create_evidence_register_sheet(wb, styles):
    """
    Create Sheet 10: Evidence_Register
    
    Purpose: Central register of supporting evidence and documentation.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Evidence_Register")
    
    # Set column widths
    widths = {
        'A': 18,  # Evidence ID
        'B': 20,  # Evidence Type
        'C': 40,  # Evidence Description
        'D': 20,  # Related Asset(s)
        'E': 20,  # Related Baseline(s)
        'F': 15,  # Evidence Date
        'G': 40,  # Evidence Location
        'H': 20,  # Evidence Owner
        'I': 15,  # Evidence Classification
        'J': 15,  # Retention Period
        'K': 15,  # Last Verified Date
        'L': 18,  # Verification Status
        'M': 25,  # Linked Control Requirement
        'N': 35   # Notes
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:N1')
    ws['A1'] = "Evidence Register - Baseline Configuration Assessment"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Evidence ID',
        'B': 'Evidence Type',
        'C': 'Evidence Description',
        'D': 'Related Asset(s)',
        'E': 'Related Baseline(s)',
        'F': 'Evidence Date',
        'G': 'Evidence Location',
        'H': 'Evidence Owner',
        'I': 'Evidence Classification',
        'J': 'Retention Period',
        'K': 'Last Verified Date',
        'L': 'Verification Status',
        'M': 'Linked Control Requirement',
        'N': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.row_dimensions[2].height = 30
    
    # Create 100 data rows
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            apply_style(cell, styles['data_cell'])
    
    # Data validations
    evidence_type_dv = create_data_validation(EVIDENCE_TYPE, allow_blank=False)
    ws.add_data_validation(evidence_type_dv)
    evidence_type_dv.add(f'B3:B{2+num_rows}')
    
    evidence_classification_dv = create_data_validation(EVIDENCE_CLASSIFICATION, allow_blank=False)
    ws.add_data_validation(evidence_classification_dv)
    evidence_classification_dv.add(f'I3:I{2+num_rows}')
    
    retention_period_dv = create_data_validation(RETENTION_PERIOD, allow_blank=False)
    ws.add_data_validation(retention_period_dv)
    retention_period_dv.add(f'J3:J{2+num_rows}')
    
    verification_status_dv = create_data_validation(VERIFICATION_STATUS, allow_blank=False)
    ws.add_data_validation(verification_status_dv)
    verification_status_dv.add(f'L3:L{2+num_rows}')
    
    # Conditional formatting - Verification Status
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Verified"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Needs Verification"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Missing"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    # Conditional formatting - Evidence Classification
    ws.conditional_formatting.add(f'I3:I{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Restricted"'], 
                   font=Font(color='9C0006')))
    ws.conditional_formatting.add(f'I3:I{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Confidential"'], 
                   font=Font(color='C65911')))
    
    # Freeze panes
    ws.freeze_panes = 'B3'
    
    return ws

def create_approval_signoff_sheet(wb, styles):
    """
    Create Sheet 11: Approval_Sign_Off
    
    Purpose: Three-tier approval signatures for the assessment.
    
    Args:
        wb: Workbook object
        styles: Style dictionary
    """
    ws = wb.create_sheet("Approval_Sign_Off")
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "Assessment Approval Sign-Off"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    row = 3
    
    # SECTION A: Document Information
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "DOCUMENT INFORMATION"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    doc_info = [
        ("Assessment Title", "Baseline Configuration Assessment - Control A.8.9"),
        ("Assessment Period", "[Start Date DD.MM.YYYY] to [End Date DD.MM.YYYY]"),
        ("Document ID", DOCUMENT_ID),
        ("Version", WORKBOOK_VERSION),
        ("Assessment Date", datetime.now().strftime('%d.%m.%Y')),
    ]
    
    row += 1
    for field, value in doc_info:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION B: Preparer Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "PREPARER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    preparer_fields = [
        ("Preparer Name", ""),
        ("Preparer Role", ""),
        ("Preparer Signature", ""),
        ("Date Prepared", ""),
        ("Completeness Attestation", 
         "I attest that the information in this assessment has been compiled accurately to the best of my knowledge and all required evidence has been collected."),
    ]
    
    row += 1
    for field, default_value in preparer_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION C: Reviewer Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "REVIEWER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    reviewer_fields = [
        ("Reviewer Name", ""),
        ("Reviewer Role", ""),
        ("Reviewer Signature", ""),
        ("Date Reviewed", ""),
        ("Review Findings", ""),
        ("Gaps Identified", ""),
        ("Review Attestation", 
         "I have reviewed this assessment and verify that it accurately represents [Organization]'s baseline configuration status. Identified gaps have been documented for remediation."),
    ]
    
    row += 1
    for field, default_value in reviewer_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
            if field in ["Review Findings", "Gaps Identified"]:
                ws.row_dimensions[row].height = 30
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        row += 1
    
    # SECTION D: Approver Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "APPROVER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    approver_fields = [
        ("Approver Name", ""),
        ("Approver Role", ""),
        ("Approver Signature", ""),
        ("Date Approved", ""),
        ("Approval Decision", ""),  # Dropdown
        ("Conditions/Comments", ""),
        ("Next Assessment Due", ""),
        ("Approver Attestation", 
         "I approve this baseline configuration assessment and authorize any documented remediation activities to proceed. This assessment will be used for compliance reporting and audit purposes."),
    ]
    
    row += 1
    for field, default_value in approver_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
            if field == "Conditions/Comments":
                ws.row_dimensions[row].height = 30
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Add dropdown for Approval Decision
        if field == "Approval Decision":
            approval_decision_dv = create_data_validation(APPROVAL_DECISION, allow_blank=True)
            ws.add_data_validation(approval_decision_dv)
            approval_decision_dv.add(f'B{row}')
        
        row += 1
    
    # Conditional formatting for Approval Decision cell
    # Find the row where Approval Decision is
    for r in range(3, row):
        if ws[f'A{r}'].value == "Approval Decision":
            ws.conditional_formatting.add(f'B{r}:B{r}',
                CellIsRule(operator='equal', formula=['"Approved"'], 
                           fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
            ws.conditional_formatting.add(f'B{r}:B{r}',
                CellIsRule(operator='equal', formula=['"Approved with Conditions"'], 
                           fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
            ws.conditional_formatting.add(f'B{r}:B{r}',
                CellIsRule(operator='equal', formula=['"Not Approved - Revisions Required"'], 
                           fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
            break
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """
    Main function to generate the baseline configuration assessment workbook.
    """
    logger.info("=" * 70)
    logger.info(f"Generating {WORKBOOK_TITLE} Workbook")
    logger.info("=" * 70)
    logger.info(f"Document ID: {DOCUMENT_ID}")
    logger.info(f"Version: {WORKBOOK_VERSION}")
    logger.info(f"Date: {datetime.now().strftime('%d.%m.%Y')}")
    logger.info("-" * 70)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create styles
    styles = create_styles()
    
    # Create all sheets
    logger.info("Creating sheets...")
    
    logger.info("  1/12 Creating Lookup Tables (hidden)...")
    create_lookup_tables(wb, styles)
    
    logger.info("  2/12 Creating Instructions sheet...")
    create_instructions_sheet(wb, styles)
    
    logger.info("  3/12 Creating Asset_Inventory sheet (100 rows)...")
    create_asset_inventory_sheet(wb, styles)
    
    logger.info("  4/12 Creating Baseline_Repository sheet (50 rows)...")
    create_baseline_repository_sheet(wb, styles)
    
    logger.info("  5/12 Creating Baseline_Coverage_Matrix sheet (43 asset types)...")
    create_baseline_coverage_matrix_sheet(wb, styles)
    
    logger.info("  6/12 Creating Approval_Tracking sheet (50 rows)...")
    create_approval_tracking_sheet(wb, styles)
    
    logger.info("  7/12 Creating Documentation_Assessment sheet (30 rows)...")
    create_documentation_assessment_sheet(wb, styles)
    
    logger.info("  8/12 Creating Version_Control sheet (50 rows)...")
    create_version_control_sheet(wb, styles)
    
    logger.info("  9/12 Creating Deviation_Register sheet (50 rows)...")
    create_deviation_register_sheet(wb, styles)
    
    logger.info(" 10/12 Creating Metrics_Summary sheet (dashboard)...")
    create_metrics_summary_sheet(wb, styles)
    
    logger.info(" 11/12 Creating Evidence_Register sheet (100 rows)...")
    create_evidence_register_sheet(wb, styles)
    
    logger.info(" 12/12 Creating Approval_Sign_Off sheet...")
    create_approval_signoff_sheet(wb, styles)
    
    logger.info("  ✓ All sheets created successfully")
    
    # Set workbook properties
    wb.properties.title = WORKBOOK_TITLE
    wb.properties.subject = f"ISMS Control A.8.9 - {WORKBOOK_TITLE}"
    wb.properties.creator = "[Organization] ISMS Implementation Team"
    wb.properties.created = datetime.now()
    wb.properties.description = "Assessment workbook for ISO 27001:2022 Control A.8.9 Baseline Configuration requirements"
    
    # Save workbook
    logger.info("-" * 70)
    logger.info("Saving workbook...")
    wb.save(FILENAME)
    
    logger.info("=" * 70)
    logger.info("✓ Workbook generated successfully!")
    logger.info("=" * 70)
    print(f"Output File: {FILENAME}")  # Changed from output_path
    logger.info(f"File Size: {os.path.getsize(FILENAME) / 1024:.1f} KB")  
    logger.info(f"Total Sheets: 12 (11 visible + 1 hidden lookup table)")
    logger.info("-" * 70)
    logger.info("\nWorkbook Structure:")
    logger.info("  1.  Instructions - Usage guidance and legend")
    logger.info("  2.  Asset_Inventory - 100 rows for asset baseline tracking")
    logger.info("  3.  Baseline_Repository - 50 rows for baseline catalog")
    logger.info("  4.  Baseline_Coverage_Matrix - 43 asset types coverage analysis")
    logger.info("  5.  Approval_Tracking - 50 rows for approval workflow")
    logger.info("  6.  Documentation_Assessment - 30 rows for quality evaluation")
    logger.info("  7.  Version_Control - 50 rows for version history")
    logger.info("  8.  Deviation_Register - 50 rows for authorized deviations")
    logger.info("  9.  Metrics_Summary - Auto-calculated compliance dashboard")
    logger.info("  10. Evidence_Register - 100 rows for evidence documentation")
    logger.info("  11. Approval_Sign_Off - Three-tier approval signatures")
    logger.info("  12. Lookup_Tables (hidden) - 43-type asset taxonomy")
    logger.info("-" * 70)
    logger.info("\nNext Steps:")
    logger.info("1. Open the workbook in Excel/LibreOffice")
    logger.info("2. Verify all sheets, data validations, and formulas")
    logger.info("3. Review Instructions sheet for usage guidance")
    logger.info("4. Customize dropdown values if needed (see CONFIGURATION section)")
    logger.info("5. Distribute to stakeholders for completion")
    logger.info("6. Use Metrics_Summary sheet for executive reporting")
    logger.info("-" * 70)
    logger.info("\nIMPORTANT REMINDERS:")
    logger.info("\u2022 This is a SAMPLE workbook - customize for your organization")
    logger.info("\u2022 Review all '# CUSTOMIZE:' comments in the script")
    logger.info("\u2022 Protected cells (gray) contain formulas - do not edit")
    logger.info("\u2022 Use dropdowns for standardized data entry")
    logger.info("\u2022 Complete Approval_Sign_Off sheet last")
    logger.info("\u2022 Retain workbook and evidence for audit (minimum 3 years)")
    logger.info("=" * 70)

if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
