#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.15.3 - Log Protection & Retention Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.15: Logging
Assessment Domain 3 of 4: Log Protection, Integrity, and Retention Management

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific log protection mechanisms, retention requirements,
and compliance obligations.

Key customization areas:
1. Log protection mechanisms (match your security controls)
2. Retention periods by log type (specific to your regulatory requirements)
3. Storage architecture and capacity (based on your infrastructure)
4. Integrity verification methods (adapt to your technical capabilities)
5. Compliance scoring criteria (aligned with your policy requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.15 Logging Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
log protection controls, integrity mechanisms, and retention management across
the organization's logging infrastructure.

**Purpose:**
Enables systematic assessment of log security controls against ISO 27001:2022
Control A.8.15 requirements, supporting evidence-based validation of log
protection, tamper detection, and retention compliance.

**Assessment Scope:**
- Log file access controls and permissions
- Log tampering protection mechanisms
- Cryptographic integrity verification (hashing, signing)
- Write-once-read-many (WORM) storage for critical logs
- Retention period management by log type
- Storage capacity and lifecycle management
- Archive and long-term storage mechanisms
- Secure deletion procedures for expired logs
- Backup and disaster recovery for log data
- Gap analysis for inadequate protection
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and protection standards
2. Access Control Assessment - Log file permission evaluation
3. Tamper Protection - Write protection and integrity controls
4. Integrity Verification - Hashing and cryptographic verification
5. Retention Management - Retention period compliance by log type
6. Storage Architecture - Capacity, tiering, and lifecycle management
7. Archive Management - Long-term storage and retrieval capabilities
8. Secure Deletion - End-of-lifecycle secure disposal procedures
9. Backup & Recovery - Log data protection and disaster recovery
10. Gap Analysis - Logs without adequate protection or retention
11. Evidence Register - Audit evidence tracking and documentation
12. Summary Dashboard - Protection and retention compliance metrics
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dropdown lists for consistency
- Conditional formatting for protection status visualization
- Automated gap identification for inadequate controls
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with log source and collection assessments

**Integration:**
This assessment feeds into the A.8.15.5 Compliance Dashboard, which
consolidates data from all four logging assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a815_3_log_protection_retention.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a815_3_log_protection_retention.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a815_3_log_protection_retention.py --date 20250124

Output:
    File: ISMS_A_8_15_3_Log_Protection_Retention_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review retention requirements from policy and regulations
    2. Document log protection mechanisms for each log type
    3. Complete protection assessments for all log sources
    4. Validate retention period compliance by log category
    5. Review storage capacity and growth projections
    6. Assess integrity verification and tamper protection
    7. Define remediation actions for inadequate protection
    8. Collect and link audit evidence (configs, retention reports)
    9. Obtain stakeholder approvals
    10. Feed results into A.8.15.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.15
Assessment Domain:    3 of 4 (Log Protection, Integrity, and Retention)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.15: Logging Policy (Governance)
    - ISMS-IMP-A.8.15.1: Log Source Inventory Assessment (Domain 1)
    - ISMS-IMP-A.8.15.2: Log Collection & Centralization Assessment (Domain 2)
    - ISMS-IMP-A.8.15.3: Log Protection & Retention Implementation Guide
    - ISMS-IMP-A.8.15.4: Log Analysis & Review Assessment (Domain 4)
    - ISMS-IMP-A.8.15.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.15.3 specification
    - Supports comprehensive log protection and retention evaluation
    - Integrated with A.8.15.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Log Protection Criticality:**
Log integrity is fundamental to incident response and forensic analysis.
Tampering with logs can conceal security breaches and undermine investigations.
Implement strong protection controls (write-once storage, cryptographic signing)
for security-relevant and compliance-mandated logs.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of log protection mechanisms and demonstration
of retention period compliance, particularly for security events and access logs.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Log storage architecture and capacity
- Security control implementation details
- Retention period mappings to regulatory requirements
- Protection mechanism configurations

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Monitor storage capacity utilization and growth rates
- Quarterly: Validate retention period compliance for all log types
- Semi-annually: Review retention requirements for regulatory changes
- Annually: Complete reassessment of protection mechanisms
- Ad-hoc: When regulations change or protection failures detected

**Quality Assurance:**
Have information security officers and compliance personnel validate retention
period mappings. Have SIEM/storage administrators validate technical protection
mechanisms before using results for compliance reporting.

**Regulatory Retention Requirements:**
Common retention periods by regulation:
- PCI DSS v4.0.1: 1 year online + 3 months archive minimum
- GDPR: Varies by processing purpose (typically 6 months to 3 years)
- HIPAA: 6 years minimum
- SOX: 7 years for financial audit logs
- Industry-specific: Consult legal/compliance for your requirements

Document the legal/regulatory basis for each retention period in policy.

**Storage Capacity Planning:**
Calculate storage requirements based on:
- Average daily log volume per source
- Retention periods by log type
- Compression ratios (typically 5:1 to 10:1)
- Growth rate (typically 15-25% annual increase)

Plan for 150% of calculated capacity to accommodate unexpected growth.

**WORM Storage Considerations:**
For critical logs requiring tamper-evident storage:
- Use WORM-capable storage (hardware or software-based)
- Implement file-level or object-level immutability
- Validate write-once enforcement through testing
- Document retention lock periods aligned with requirements

**Chain of Custody:**
For logs used in legal/forensic contexts, maintain chain of custody:
- Document access to protected logs
- Use cryptographic hashing to prove non-modification
- Implement secure transfer procedures for evidence delivery
- Retain metadata proving protection mechanism effectiveness

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

# Document identification constants
DOCUMENT_ID = "ISMS-IMP-A.8.15.3"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.8.15: Logging"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Log_Protection_Retention_{GENERATED_TIMESTAMP}.xlsx"


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLES
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 15 sheets - comprehensive log protection & retention assessment
    sheets = [
        "Instructions & Legend",
        "Access Control Assessment",
        "Integrity Protection",
        "Secure Transmission",
        "Retention Period Compliance",
        "Storage Tier Implementation",
        "Log Backup & Recovery",
        "Disposal Procedures",
        "Separation of Duties",
        "Legal Hold Management",
        "Privacy Impact Assessment",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval & Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def setup_styles():
    """Define all cell styles. Same as A.8.15.1 and A.8.15.2 for consistency."""
    return {
        'header_main': {
            'font': {'name': 'Calibri', 'size': 14, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '003366', 'end_color': '003366', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'header_sub': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '4472C4', 'end_color': '4472C4', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'column_header': {
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '000000'},
            'fill': {'start_color': 'D9D9D9', 'end_color': 'D9D9D9', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'input_cell': {
            'fill': {'start_color': 'FFFFCC', 'end_color': 'FFFFCC', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'example_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'italic': True, 'color': '666666'},
            'alignment': {'horizontal': 'left', 'vertical': 'top'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'formula_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'info_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'section_header': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '4472C4', 'end_color': '4472C4', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        }
    }


def apply_style(cell, style_template):
    """Apply a style template to a cell."""
    if 'font' in style_template:
        cell.font = Font(**style_template['font'])
    if 'fill' in style_template:
        cell.fill = PatternFill(**style_template['fill'])
    if 'alignment' in style_template:
        cell.alignment = Alignment(**style_template['alignment'])
    if 'border' in style_template:
        cell.border = Border(
            left=Side(style=style_template['border'].get('left', 'thin')),
            right=Side(style=style_template['border'].get('right', 'thin')),
            top=Side(style=style_template['border'].get('top', 'thin')),
            bottom=Side(style=style_template['border'].get('bottom', 'thin'))
        )


def set_column_widths(ws, widths):
    """Set column widths."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ============================================================================
# SECTION 2: INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # Main header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{DOCUMENT_ID}  -  Log Protection & Retention Assessment\n{CONTROL_REF}"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    row = 4
    info_fields = [
        ("Document ID:", "ISMS-IMP-A.8.15.3"),
        ("Assessment Area:", "Log Protection, Integrity, and Retention Compliance"),
        ("Related Policy:", "ISMS-POL-A.8.15-S2.2, S2.3"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT - DD.MM.YYYY]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Semi-annual"),
    ]
    
    for label, value in info_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = value
        
        if "[USER INPUT" in value:
            apply_style(ws[f'B{row}'], styles['input_cell'])
            if "Date" in label:
                ws[f'B{row}'].number_format = 'DD.MM.YYYY'
        else:
            apply_style(ws[f'B{row}'], styles['info_cell'])
        
        row += 1
    
    # Key definitions
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "KEY DEFINITIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    definitions = [
        ("WORM", "Write-Once-Read-Many (immutable storage)"),
        ("Cryptographic Hashing", "SHA-256/SHA-512 for integrity verification"),
        ("Digital Signature", "Cryptographic proof of authenticity"),
        ("Hot/Warm/Cold Storage", "Tiered retention based on access frequency"),
        ("Legal Hold", "Suspension of normal retention/disposal for litigation"),
        ("Separation of Duties", "System admins cannot modify their own logs"),
        ("Secure Disposal", "Cryptographic erasure or physical destruction"),
    ]
    
    for term, definition in definitions:
        ws[f'A{row}'] = term
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = definition
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
    
    set_column_widths(ws, {'A': 25, 'B': 60, 'C': 15, 'D': 15, 'E': 15, 'F': 15})
    ws.freeze_panes = 'A3'


# ============================================================================
# SECTION 3: ACCESS CONTROL ASSESSMENT SHEET
# ============================================================================

def create_access_control_sheet(ws, styles):
    """
    Create Access Control Assessment sheet.
    
    "Security is a process, not a product." - Bruce Schneier
    Let's assess the process of controlling log access!
    """
    
    ws.merge_cells('A1:P1')
    ws['A1'] = "ACCESS CONTROL ASSESSMENT"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:P2')
    ws['A2'] = "Assess log access controls per ISMS-POL-A.8.15-S2.2.2"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Log Source / SIEM Component", 30),
        ("B", "Access Control Type", 20),
        ("C", "Authentication Required", 18),
        ("D", "Authorization Model", 20),
        ("E", "Read Access Controlled", 18),
        ("F", "Write Access Prevented", 20),
        ("G", "Delete Access Controlled", 20),
        ("H", "Admin Separation", 20),
        ("I", "Access Logged (Meta)", 20),
        ("J", "MFA Required for Admin", 20),
        ("K", "Last Access Review", 18),
        ("L", "Review Frequency", 20),
        ("M", "Non-Compliance Issues", 40),
        ("N", "Compliance Score", 15),
        ("O", "Remediation Required", 18),
        ("P", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "SIEM Core", "RBAC", "Yes", "Role-Based", "Yes", "Yes (read-only)",
        "Yes (restricted)", "Yes (separated)", "Yes", "Yes", "15.12.2025",
        "Quarterly", "None", "100%", "No", "Full compliance"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 101):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'O', 'P']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column K: Date format
        apply_style(ws[f'K{data_row}'], styles['input_cell'])
        ws[f'K{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column N: Compliance Score Formula
        ws[f'N{data_row}'] = f'=IF(A{data_row}="","",(COUNTIF(C{data_row}:J{data_row},"*Yes*")+COUNTIF(C{data_row}:J{data_row},"Yes*"))/8)'
        ws[f'N{data_row}'].number_format = '0.0%'
        apply_style(ws[f'N{data_row}'], styles['formula_cell'])
    
    # Data validations
    access_type_dv = DataValidation(type="list",
        formula1='"RBAC,ACL,None,Other"', allow_blank=True)
    ws.add_data_validation(access_type_dv)
    access_type_dv.add('B9:B100')
    
    yes_no_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('C9:C100')
    yes_no_dv.add('O9:O100')
    
    auth_model_dv = DataValidation(type="list",
        formula1='"Role-Based,User-Based,Group-Based,None"', allow_blank=True)
    ws.add_data_validation(auth_model_dv)
    auth_model_dv.add('D9:D100')
    
    read_dv = DataValidation(type="list",
        formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(read_dv)
    read_dv.add('E9:E100')
    
    write_dv = DataValidation(type="list",
        formula1='"Yes (read-only),No,Partial"', allow_blank=True)
    ws.add_data_validation(write_dv)
    write_dv.add('F9:F100')
    
    delete_dv = DataValidation(type="list",
        formula1='"Yes (restricted),No"', allow_blank=True)
    ws.add_data_validation(delete_dv)
    delete_dv.add('G9:G100')
    
    admin_sep_dv = DataValidation(type="list",
        formula1='"Yes (separated),No,N/A"', allow_blank=True)
    ws.add_data_validation(admin_sep_dv)
    admin_sep_dv.add('H9:H100')
    
    logged_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(logged_dv)
    logged_dv.add('I9:I100')
    
    mfa_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,➖ N/A"', allow_blank=True)
    ws.add_data_validation(mfa_dv)
    mfa_dv.add('J9:J100')
    
    frequency_dv = DataValidation(type="list",
        formula1='"Quarterly,Semi-annual,Annual,None"', allow_blank=True)
    ws.add_data_validation(frequency_dv)
    frequency_dv.add('L9:L100')
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 4: INTEGRITY PROTECTION MECHANISMS SHEET
# ============================================================================

def create_integrity_protection_sheet(ws, styles):
    """Create Integrity Protection Mechanisms sheet."""
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "INTEGRITY PROTECTION MECHANISMS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:Q2')
    ws['A2'] = "Assess integrity protection per ISMS-POL-A.8.15-S2.2.4"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Log Source / Storage", 30),
        ("B", "Log Criticality", 18),
        ("C", "Write-Once Storage (WORM)", 20),
        ("D", "WORM Technology", 25),
        ("E", "Cryptographic Hashing", 20),
        ("F", "Hash Algorithm", 18),
        ("G", "Hash Storage Location", 25),
        ("H", "Digital Signatures", 18),
        ("I", "File Sealing", 18),
        ("J", "Integrity Check Frequency", 20),
        ("K", "Last Integrity Check", 18),
        ("L", "Tampering Detected", 18),
        ("M", "Backup Protected", 18),
        ("N", "Compliance with Policy", 20),
        ("O", "Gap Description", 40),
        ("P", "Remediation Plan", 40),
        ("Q", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Security Logs - Hot Storage", "Critical", "Yes", "Cloud Object Lock",
        "Yes", "SHA-256", "Separate integrity DB", "No", "N/A", "Daily",
        "06.01.2026", "Never", "Yes", "Compliant", "", "", "Full protection"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 101):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'O', 'P', 'Q']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column K: Date format
        apply_style(ws[f'K{data_row}'], styles['input_cell'])
        ws[f'K{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column N: Compliance Formula
        # Critical: WORM + Hash + Daily checks
        # High: Hash + Weekly checks
        # Medium: Access controls + Monthly checks
        ws[f'N{data_row}'] = f'''=IF(A{data_row}="","",
            IF(B{data_row}="Critical",
                IF(AND(C{data_row}="Yes",E{data_row}="Yes",J{data_row}="Daily"),"Compliant","Non-Compliant"),
            IF(B{data_row}="High",
                IF(AND(E{data_row}="Yes",OR(J{data_row}="Daily",J{data_row}="Weekly")),"Compliant","Non-Compliant"),
            IF(B{data_row}="Medium",
                IF(OR(J{data_row}="Daily",J{data_row}="Weekly",J{data_row}="Monthly"),"Compliant","Non-Compliant"),
                "Compliant"))))'''
        apply_style(ws[f'N{data_row}'], styles['formula_cell'])
    
    # Data validations
    criticality_dv = DataValidation(type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"', allow_blank=True)
    ws.add_data_validation(criticality_dv)
    criticality_dv.add('B9:B100')
    
    yes_no_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,➖ N/A"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('C9:C100')
    yes_no_dv.add('E9:E100')
    yes_no_dv.add('H9:H100')
    yes_no_dv.add('I9:I100')
    yes_no_dv.add('M9:M100')
    
    worm_tech_dv = DataValidation(type="list",
        formula1='"Hardware WORM,Software WORM,Cloud Object Lock,None"', allow_blank=True)
    ws.add_data_validation(worm_tech_dv)
    worm_tech_dv.add('D9:D100')
    
    hash_algo_dv = DataValidation(type="list",
        formula1='"SHA-256,SHA-512,SHA-3,MD5 (weak),None"', allow_blank=True)
    ws.add_data_validation(hash_algo_dv)
    hash_algo_dv.add('F9:F100')
    
    frequency_dv = DataValidation(type="list",
        formula1='"Daily,Weekly,Monthly,None"', allow_blank=True)
    ws.add_data_validation(frequency_dv)
    frequency_dv.add('J9:J100')
    
    tampering_dv = DataValidation(type="list",
        formula1='"Never,Historical,Recent"', allow_blank=True)
    ws.add_data_validation(tampering_dv)
    tampering_dv.add('L9:L100')
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 5: SECURE TRANSMISSION ASSESSMENT SHEET
# ============================================================================

def create_secure_transmission_sheet(ws, styles):
    """Create Secure Transmission Assessment sheet."""
    
    ws.merge_cells('A1:N1')
    ws['A1'] = "SECURE TRANSMISSION ASSESSMENT"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Assess log transmission security per ISMS-POL-A.8.15-S2.2.3"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Source System", 30),
        ("B", "Destination (SIEM)", 25),
        ("C", "Transport Protocol", 20),
        ("D", "Encryption in Transit", 18),
        ("E", "TLS Version", 15),
        ("F", "Certificate Validation", 20),
        ("G", "Network Segment", 20),
        ("H", "Firewall Protection", 18),
        ("I", "Source Authentication", 20),
        ("J", "Vulnerability Risk", 18),
        ("K", "Compliance Status", 18),
        ("L", "Remediation Required", 18),
        ("M", "Target Date", 15),
        ("N", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "web-prod-01", "siem-core-01", "TLS", "Yes (TLS)", "TLS 1.3",
        "Yes", "Internal Network", "Yes", "Yes (mutual TLS/certs)",
        "None", "Compliant", "No", "", "Secure configuration"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 201):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'N']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column K: Compliance Status Formula
        # Compliant if TLS 1.2+ on untrusted networks (Internet, DMZ)
        ws[f'K{data_row}'] = f'''=IF(A{data_row}="","",
            IF(OR(G{data_row}="Internet",G{data_row}="DMZ"),
                IF(OR(E{data_row}="TLS 1.3",E{data_row}="TLS 1.2"),"Compliant","Non-Compliant"),
                IF(D{data_row}="Yes (TLS)","Compliant","Partial")))'''
        apply_style(ws[f'K{data_row}'], styles['formula_cell'])
        
        # Column M: Target Date
        apply_style(ws[f'M{data_row}'], styles['input_cell'])
        ws[f'M{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    protocol_dv = DataValidation(type="list",
        formula1='"TLS,TCP,UDP,HTTPS,Other"', allow_blank=True)
    ws.add_data_validation(protocol_dv)
    protocol_dv.add('C9:C200')
    
    encryption_dv = DataValidation(type="list",
        formula1='"Yes (TLS),No"', allow_blank=True)
    ws.add_data_validation(encryption_dv)
    encryption_dv.add('D9:D200')
    
    tls_version_dv = DataValidation(type="list",
        formula1='"TLS 1.3,TLS 1.2,TLS 1.1 (weak),TLS 1.0 (weak),None"', allow_blank=True)
    ws.add_data_validation(tls_version_dv)
    tls_version_dv.add('E9:E200')
    
    cert_val_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,➖ N/A"', allow_blank=True)
    ws.add_data_validation(cert_val_dv)
    cert_val_dv.add('F9:F200')
    
    network_dv = DataValidation(type="list",
        formula1='"Isolated Mgmt Network,Internal Network,Internet,DMZ"', allow_blank=True)
    ws.add_data_validation(network_dv)
    network_dv.add('G9:G200')
    
    yes_no_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('H9:H200')
    yes_no_dv.add('L9:L200')
    
    auth_dv = DataValidation(type="list",
        formula1='"Yes (mutual TLS/certs),No"', allow_blank=True)
    ws.add_data_validation(auth_dv)
    auth_dv.add('I9:I200')
    
    risk_dv = DataValidation(type="list",
        formula1='"None,Low,Medium,High"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add('J9:J200')
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 6: RETENTION PERIOD COMPLIANCE SHEET
# ============================================================================

def create_retention_period_sheet(ws, styles):
    """Create Retention Period Compliance sheet."""
    
    ws.merge_cells('A1:R1')
    ws['A1'] = "RETENTION PERIOD COMPLIANCE"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:R2')
    ws['A2'] = "Assess retention period compliance per ISMS-POL-A.8.15-S2.3.2"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Log Source / Type", 30),
        ("B", "Log Category", 20),
        ("C", "Regulatory Requirement", 25),
        ("D", "Policy Retention (months)", 20),
        ("E", "Hot Storage Period (months)", 22),
        ("F", "Warm Storage Period (months)", 22),
        ("G", "Cold Storage Period (years)", 20),
        ("H", "Total Retention (months)", 20),
        ("I", "Meets Policy Requirement", 20),
        ("J", "Retention Gap (months)", 20),
        ("K", "Over-Retention (months)", 20),
        ("L", "Automated Disposal", 18),
        ("M", "Last Disposal Date", 18),
        ("N", "Legal Hold Capability", 18),
        ("O", "Compliance Status", 18),
        ("P", "Remediation Action", 40),
        ("Q", "Target Date", 15),
        ("R", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Authentication Logs", "Authentication", "ISO 27001", "12", "3", "9",
        "0", "12", "Yes", "0", "0", "Yes", "01.01.2026", "Yes",
        "Compliant", "", "", "Meets minimum retention"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 101):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'L', 'N', 'P', 'R']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column H: Total Retention Formula
        ws[f'H{data_row}'] = f'=IF(A{data_row}="","",E{data_row}+F{data_row}+(G{data_row}*12))'
        apply_style(ws[f'H{data_row}'], styles['formula_cell'])
        
        # Column I: Meets Policy Formula
        ws[f'I{data_row}'] = f'=IF(A{data_row}="","",IF(H{data_row}>=D{data_row},"Yes","No"))'
        apply_style(ws[f'I{data_row}'], styles['formula_cell'])
        
        # Column J: Retention Gap
        ws[f'J{data_row}'] = f'=IF(I{data_row}="Yes",0,D{data_row}-H{data_row})'
        apply_style(ws[f'J{data_row}'], styles['formula_cell'])
        
        # Column K: Over-Retention
        ws[f'K{data_row}'] = f'=IF(H{data_row}<=D{data_row}*1.5,0,H{data_row}-D{data_row})'
        apply_style(ws[f'K{data_row}'], styles['formula_cell'])
        
        # Column M: Date format
        apply_style(ws[f'M{data_row}'], styles['input_cell'])
        ws[f'M{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column O: Compliance Status
        ws[f'O{data_row}'] = f'=IF(A{data_row}="","",IF(I{data_row}="Yes","Compliant","Non-Compliant"))'
        apply_style(ws[f'O{data_row}'], styles['formula_cell'])
        
        # Column Q: Target Date
        apply_style(ws[f'Q{data_row}'], styles['input_cell'])
        ws[f'Q{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    category_dv = DataValidation(type="list",
        formula1='"Security,Authentication,Admin,Application,System,Network,Database"', allow_blank=True)
    ws.add_data_validation(category_dv)
    category_dv.add('B9:B100')
    
    regulatory_dv = DataValidation(type="list",
        formula1='"PCI DSS v4.0.1,HIPAA,SOX,GDPR,FADP,ISO 27001,None"', allow_blank=True)
    ws.add_data_validation(regulatory_dv)
    regulatory_dv.add('C9:C100')
    
    yes_no_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('L9:L100')
    yes_no_dv.add('N9:N100')
    
    # Quick Reference Table
    row = 105
    ws.merge_cells(f'A{row}:R{row}')
    ws[f'A{row}'] = "QUICK REFERENCE - STANDARD RETENTION PERIODS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    ref_headers = ["Log Category", "Minimum Retention", "Regulatory Notes"]
    for col_idx, header in enumerate(ref_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    reference_data = [
        ("Security Events", "12 months", "ISO 27001, most regulations"),
        ("Authentication", "12 months", "PCI DSS v4.0.1: 12 months minimum"),
        ("Administrative Actions", "12 months", "SOX: audit trail required"),
        ("Financial Transactions", "7 years", "SOX compliance"),
        ("Healthcare Access", "6 years", "HIPAA requirements"),
        ("Payment Card Data", "12 months", "PCI DSS v4.0.1"),
    ]
    
    for category, retention, notes in reference_data:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = retention
        ws[f'C{row}'] = notes
        ws.merge_cells(f'C{row}:R{row}')
        row += 1
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 7: STORAGE TIER IMPLEMENTATION SHEET
# ============================================================================

def create_storage_tier_sheet(ws, styles):
    """Create Storage Tier Implementation sheet."""
    
    ws.merge_cells('A1:O1')
    ws['A1'] = "STORAGE TIER IMPLEMENTATION"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:O2')
    ws['A2'] = "Assess tiered storage implementation per ISMS-POL-A.8.15-S2.3.3"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Storage Tier", 20),
        ("B", "Technology", 25),
        ("C", "Capacity (TB)", 15),
        ("D", "Used (TB)", 15),
        ("E", "% Used", 12),
        ("F", "Retention Period", 20),
        ("G", "Access Performance", 20),
        ("H", "Encryption at Rest", 18),
        ("I", "Encryption Method", 20),
        ("J", "Geographic Location", 25),
        ("K", "Redundancy", 20),
        ("L", "Backup Implemented", 18),
        ("M", "Meets Policy Requirements", 20),
        ("N", "Issues", 40),
        ("O", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Hot", "Local SSD", "10", "6.5", "65%", "0-90 days",
        "Real-time (<1 min)", "Yes", "AES-256", "Primary DC",
        "Both", "Yes", "Yes", "None", "Production hot storage"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 31):
        for col_letter in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column E: % Used Formula
        ws[f'E{data_row}'] = f'=IF(C{data_row}=0,0,D{data_row}/C{data_row})'
        ws[f'E{data_row}'].number_format = '0.0%'
        apply_style(ws[f'E{data_row}'], styles['formula_cell'])
    
    # Data validations
    tier_dv = DataValidation(type="list",
        formula1='"Hot,Warm,Cold"', allow_blank=True)
    ws.add_data_validation(tier_dv)
    tier_dv.add('A9:A30')
    
    tech_dv = DataValidation(type="list",
        formula1='"Local Disk/SSD,SAN,NAS,Object Storage,Tape,Other"', allow_blank=True)
    ws.add_data_validation(tech_dv)
    tech_dv.add('B9:B30')
    
    performance_dv = DataValidation(type="list",
        formula1='"Real-time (<1 min),Fast (<15 min),Slow (hours),Very Slow (days)"', allow_blank=True)
    ws.add_data_validation(performance_dv)
    performance_dv.add('G9:G30')
    
    yes_no_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('H9:H30')
    yes_no_dv.add('L9:L30')
    yes_no_dv.add('M9:M30')
    
    encryption_dv = DataValidation(type="list",
        formula1='"AES-256,AES-128,None"', allow_blank=True)
    ws.add_data_validation(encryption_dv)
    encryption_dv.add('I9:I30')
    
    redundancy_dv = DataValidation(type="list",
        formula1='"None,Local (RAID),Remote (Replication),Both"', allow_blank=True)
    ws.add_data_validation(redundancy_dv)
    redundancy_dv.add('K9:K30')
    
    # Tiering Assessment Questions
    row = 35
    ws.merge_cells(f'A{row}:O{row}')
    ws[f'A{row}'] = "TIERING ASSESSMENT QUESTIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    questions = [
        "☐ Is 3-tier model implemented (hot/warm/cold)?",
        "☐ Are transitions between tiers automated?",
        "☐ Are retention periods matched to storage tiers?",
        "☐ Is encryption at rest implemented for all tiers?",
        "☐ Is geographic redundancy implemented for critical logs?",
        "☐ Are backups taken from all tiers?",
        "☐ Is performance adequate for each tier's use case?",
    ]
    
    for question in questions:
        ws.merge_cells(f'A{row}:O{row}')
        ws[f'A{row}'] = question
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 8: LOG BACKUP & RECOVERY SHEET
# ============================================================================

def create_backup_recovery_sheet(ws, styles):
    """
    Create Log Backup & Recovery sheet.
    
    "Hope is not a strategy." - Ancient Sysadmin Proverb
    Test your backups or be disappointed!
    """
    
    ws.merge_cells('A1:P1')
    ws['A1'] = "LOG BACKUP & RECOVERY"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:P2')
    ws['A2'] = "Assess backup and recovery capabilities per ISMS-POL-A.8.15-S2.2.7"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Backup Scope", 30),
        ("B", "Backup Frequency", 20),
        ("C", "Backup Technology", 25),
        ("D", "Backup Location", 30),
        ("E", "Backup Encrypted", 18),
        ("F", "Encryption Algorithm", 20),
        ("G", "Backup Integrity Verified", 20),
        ("H", "Last Backup Date", 18),
        ("I", "Last Restore Test Date", 18),
        ("J", "Restore Test Frequency", 20),
        ("K", "Last Restore Success", 18),
        ("L", "RTO (Recovery Time)", 20),
        ("M", "RPO (Recovery Point)", 20),
        ("N", "Backup Retention Period", 20),
        ("O", "Compliance Status", 18),
        ("P", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "All Logs", "Daily", "Veeam", "Offsite", "Yes", "AES-256",
        "Yes (always)", "06.01.2026", "15.12.2025", "Quarterly", "Yes",
        "24 hours", "24 hours", "90 days", "Compliant", "Full backup daily"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 26):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'J', 'K', 'L', 'M', 'N', 'P']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Columns H, I: Date format
        for col_letter in ['H', 'I']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
            ws[f'{col_letter}{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column O: Compliance Status Formula
        # Compliant if: Daily backups + Offsite + Encrypted + Quarterly testing
        ws[f'O{data_row}'] = f'''=IF(A{data_row}="","",
            IF(AND(B{data_row}="Daily",OR(D{data_row}="Offsite",D{data_row}="Cloud",D{data_row}="Multiple"),
                   E{data_row}="Yes",OR(J{data_row}="Quarterly",J{data_row}="Monthly")),
                "Compliant","Partial"))'''
        apply_style(ws[f'O{data_row}'], styles['formula_cell'])
    
    # Data validations
    scope_dv = DataValidation(type="list",
        formula1='"All Logs,Hot Storage Only,Critical Logs Only,None"', allow_blank=True)
    ws.add_data_validation(scope_dv)
    scope_dv.add('A9:A25')
    
    frequency_dv = DataValidation(type="list",
        formula1='"Daily,Weekly,Monthly"', allow_blank=True)
    ws.add_data_validation(frequency_dv)
    frequency_dv.add('B9:B25')
    
    location_dv = DataValidation(type="list",
        formula1='"Same Site,Offsite,Cloud,Multiple"', allow_blank=True)
    ws.add_data_validation(location_dv)
    location_dv.add('D9:D25')
    
    yes_no_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('E9:E25')
    yes_no_dv.add('K9:K25')
    
    encryption_dv = DataValidation(type="list",
        formula1='"AES-256,AES-128,None"', allow_blank=True)
    ws.add_data_validation(encryption_dv)
    encryption_dv.add('F9:F25')
    
    verified_dv = DataValidation(type="list",
        formula1='"Yes (periodic),Yes (always),No"', allow_blank=True)
    ws.add_data_validation(verified_dv)
    verified_dv.add('G9:G25')
    
    test_freq_dv = DataValidation(type="list",
        formula1='"Quarterly,Semi-annual,Annual,Never"', allow_blank=True)
    ws.add_data_validation(test_freq_dv)
    test_freq_dv.add('J9:J25')
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 9: DISPOSAL PROCEDURES SHEET
# ============================================================================

def create_disposal_procedures_sheet(ws, styles):
    """Create Disposal Procedures sheet."""
    
    ws.merge_cells('A1:N1')
    ws['A1'] = "DISPOSAL PROCEDURES"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Assess secure disposal practices per ISMS-POL-A.8.15-S2.3.5"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Log Type / Source", 30),
        ("B", "Retention Period Expired", 20),
        ("C", "Automated Disposal", 18),
        ("D", "Disposal Method", 25),
        ("E", "Disposal Approval Required", 22),
        ("F", "Legal Hold Check", 18),
        ("G", "Disposal Logged", 18),
        ("H", "Last Disposal Date", 18),
        ("I", "Volume Disposed (GB)", 20),
        ("J", "Disposal Verification", 20),
        ("K", "Compliance with Policy", 20),
        ("L", "Issues", 40),
        ("M", "Remediation", 40),
        ("N", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Web Server Logs (>12 months)", "Monthly", "Yes", "Cryptographic Erasure",
        "No (automated)", "Yes (checked)", "Yes", "01.01.2026", "250",
        "Verified", "Compliant", "None", "", "Automated monthly cleanup"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 51):
        for col_letter in ['A', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'L', 'M', 'N']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Columns B, H: Date format
        for col_letter in ['B', 'H']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
            ws[f'{col_letter}{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column K: Compliance Formula
        # Compliant if: Legal hold checked + Disposal logged + Secure method
        ws[f'K{data_row}'] = f'''=IF(A{data_row}="","",
            IF(AND(F{data_row}="Yes (checked)",G{data_row}="Yes",
                   OR(D{data_row}="Cryptographic Erasure",D{data_row}="Multi-pass Overwrite",D{data_row}="Physical Destruction")),
                "Compliant","Non-Compliant"))'''
        apply_style(ws[f'K{data_row}'], styles['formula_cell'])
    
    # Data validations
    automated_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(automated_dv)
    automated_dv.add('C9:C50')
    
    method_dv = DataValidation(type="list",
        formula1='"Cryptographic Erasure,Multi-pass Overwrite,Physical Destruction,Deletion"', allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add('D9:D50')
    
    approval_dv = DataValidation(type="list",
        formula1='"Yes (manual),No (automated),N/A"', allow_blank=True)
    ws.add_data_validation(approval_dv)
    approval_dv.add('E9:E50')
    
    legal_hold_dv = DataValidation(type="list",
        formula1='"Yes (checked),No,N/A"', allow_blank=True)
    ws.add_data_validation(legal_hold_dv)
    legal_hold_dv.add('F9:F50')
    
    logged_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(logged_dv)
    logged_dv.add('G9:G50')
    
    verified_dv = DataValidation(type="list",
        formula1='"Verified,Not Verified"', allow_blank=True)
    ws.add_data_validation(verified_dv)
    verified_dv.add('J9:J50')
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 10: SEPARATION OF DUTIES SHEET
# ============================================================================

def create_separation_of_duties_sheet(ws, styles):
    """Create Separation of Duties sheet."""
    
    ws.merge_cells('A1:N1')
    ws['A1'] = "SEPARATION OF DUTIES"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Assess separation of duties per ISMS-POL-A.8.15-S2.2.6"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "System / Component", 30),
        ("B", "System Administrator(s)", 30),
        ("C", "Log Administrator(s)", 30),
        ("D", "Roles Separated", 18),
        ("E", "Sys Admin Can Modify Logs", 25),
        ("F", "Compensating Controls", 40),
        ("G", "Break-Glass Procedure", 20),
        ("H", "Break-Glass Usage Logged", 22),
        ("I", "Independent Review", 20),
        ("J", "Last Review Date", 18),
        ("K", "Violations Detected", 18),
        ("L", "Compliance Status", 18),
        ("M", "Remediation Plan", 40),
        ("N", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "SIEM Platform", "IT Ops Team", "Security Team", "Yes (different people)",
        "No (compliant)", "", "Yes (documented)", "Yes", "Yes", "15.12.2025",
        "None", "Compliant", "", "Proper separation implemented"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 51):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'M', 'N']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column J: Date format
        apply_style(ws[f'J{data_row}'], styles['input_cell'])
        ws[f'J{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column L: Compliance Status Formula
        ws[f'L{data_row}'] = f'''=IF(A{data_row}="","",
            IF(AND(D{data_row}="Yes (different people)",
                   OR(E{data_row}="No (compliant)",E{data_row}="Limited")),
                "Compliant","Non-Compliant"))'''
        apply_style(ws[f'L{data_row}'], styles['formula_cell'])
    
    # Data validations
    separated_dv = DataValidation(type="list",
        formula1='"Yes (different people),No (same people),Partial"', allow_blank=True)
    ws.add_data_validation(separated_dv)
    separated_dv.add('D9:D50')
    
    modify_dv = DataValidation(type="list",
        formula1='"No (compliant),Yes (violation),Limited"', allow_blank=True)
    ws.add_data_validation(modify_dv)
    modify_dv.add('E9:E50')
    
    breakglass_dv = DataValidation(type="list",
        formula1='"Yes (documented),No"', allow_blank=True)
    ws.add_data_validation(breakglass_dv)
    breakglass_dv.add('G9:G50')
    
    yes_no_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,➖ N/A"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('H9:H50')
    yes_no_dv.add('I9:I50')
    
    violations_dv = DataValidation(type="list",
        formula1='"None,Historical,Recent"', allow_blank=True)
    ws.add_data_validation(violations_dv)
    violations_dv.add('K9:K50')
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 11: LEGAL HOLD MANAGEMENT SHEET
# ============================================================================

def create_legal_hold_sheet(ws, styles):
    """Create Legal Hold Management sheet."""
    
    ws.merge_cells('A1:M1')
    ws['A1'] = "LEGAL HOLD MANAGEMENT"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:M2')
    ws['A2'] = "Track legal holds per ISMS-POL-A.8.15-S2.3.6"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Hold ID", 15),
        ("B", "Hold Name / Matter", 30),
        ("C", "Initiation Date", 15),
        ("D", "Initiated By", 25),
        ("E", "Scope Description", 40),
        ("F", "Systems/Sources Affected", 30),
        ("G", "Date Range", 20),
        ("H", "Hold Status", 15),
        ("I", "Review Date", 15),
        ("J", "Disposal Prevented", 18),
        ("K", "Release Date", 15),
        ("L", "Release Authorized By", 25),
        ("M", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "HOLD-001", "Internal Investigation 2025-Q4", "15.10.2025",
        "Legal Counsel - Jane Smith", "All authentication and access logs for Finance Dept",
        "ERP System, File Servers", "01.09.2025 - 31.12.2025", "Active",
        "15.01.2026", "Yes (verified)", "", "", "Quarterly review scheduled"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 31):
        # Column A: Auto-generate Hold ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","HOLD-"&TEXT(ROW()-8,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        
        # Other columns: Input
        for col_letter in ['B', 'D', 'E', 'F', 'G', 'H', 'J', 'L', 'M']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Date columns
        for col_letter in ['C', 'I', 'K']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
            ws[f'{col_letter}{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    status_dv = DataValidation(type="list",
        formula1='"Active,Released,Suspended"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('H9:H30')
    
    prevented_dv = DataValidation(type="list",
        formula1='"Yes (verified),No"', allow_blank=True)
    ws.add_data_validation(prevented_dv)
    prevented_dv.add('J9:J30')
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 11.5: PRIVACY IMPACT ASSESSMENT SHEET
# ============================================================================

def create_privacy_impact_sheet(ws, styles):
    """Create Privacy Impact Assessment sheet for GDPR/data minimization."""

    ws.merge_cells('A1:K1')
    ws['A1'] = "PRIVACY IMPACT ASSESSMENT"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:K2')
    ws['A2'] = "Assess privacy implications of log data collection and retention"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30

    # ==================== DATA MINIMIZATION ASSESSMENT ====================
    row = 4
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "DATA MINIMIZATION ASSESSMENT"
    apply_style(ws[f'A{row}'], styles['section_header'])
    row += 1

    headers = ["Log Category", "PII/Sensitive Data Present", "Data Types", "Minimization Applied", "Justification", "Retention Period", "GDPR Article", "Compliance Status", "Review Date", "Owner", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles['column_header'])
    row += 1

    log_categories = [
        ("Authentication Logs", "", "", "", "", "", "Art. 6(1)(f)", "", "", "", ""),
        ("Access Logs", "", "", "", "", "", "Art. 6(1)(f)", "", "", "", ""),
        ("Application Logs", "", "", "", "", "", "Art. 6(1)(f)", "", "", "", ""),
        ("Network Logs", "", "", "", "", "", "Art. 6(1)(f)", "", "", "", ""),
        ("Database Audit Logs", "", "", "", "", "", "Art. 6(1)(c)", "", "", "", ""),
        ("Email Logs", "", "", "", "", "", "Art. 6(1)(f)", "", "", "", ""),
        ("HR System Logs", "", "", "", "", "", "Art. 6(1)(b)", "", "", "", ""),
    ]

    for category_data in log_categories:
        for col_idx, value in enumerate(category_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 1:
                apply_style(cell, styles['input_cell'])
        row += 1

    row += 2

    # ==================== PROHIBITED DATA CHECK ====================
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "PROHIBITED DATA CHECK"
    apply_style(ws[f'A{row}'], styles['section_header'])
    row += 1

    prohibited_headers = ["Data Type", "Logs Checked", "Found?", "Location", "Remediation Action", "Completion Date", "Verified By"]
    for col_idx, header in enumerate(prohibited_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles['column_header'])
    row += 1

    prohibited_types = [
        ("Passwords/Credentials", "", "", "", "", "", ""),
        ("Credit Card Numbers", "", "", "", "", "", ""),
        ("Social Security Numbers", "", "", "", "", "", ""),
        ("Health Information", "", "", "", "", "", ""),
        ("Biometric Data", "", "", "", "", "", ""),
        ("Racial/Ethnic Data", "", "", "", "", "", ""),
        ("Political Opinions", "", "", "", "", "", ""),
        ("Religious Beliefs", "", "", "", "", "", ""),
    ]

    for prohibited_data in prohibited_types:
        for col_idx, value in enumerate(prohibited_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 1:
                apply_style(cell, styles['input_cell'])
        row += 1

    row += 2

    # ==================== DATA SUBJECT RIGHTS ====================
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "DATA SUBJECT RIGHTS COMPLIANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    row += 1

    rights_headers = ["Right", "Applicable to Logs?", "Process Documented", "Can Be Fulfilled", "Response Time", "Evidence Ref"]
    for col_idx, header in enumerate(rights_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles['column_header'])
    row += 1

    rights = [
        ("Right of Access (Art. 15)", "", "", "", "", ""),
        ("Right to Rectification (Art. 16)", "", "", "", "", ""),
        ("Right to Erasure (Art. 17)", "", "", "", "", ""),
        ("Right to Restriction (Art. 18)", "", "", "", "", ""),
        ("Right to Data Portability (Art. 20)", "", "", "", "", ""),
    ]

    for right_data in rights:
        for col_idx, value in enumerate(right_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 1:
                apply_style(cell, styles['input_cell'])
        row += 1

    # Column widths
    widths = [25, 22, 20, 20, 25, 18, 15, 18, 14, 15, 30]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A6'


# ============================================================================
# SECTION 12: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create Gap Analysis sheet."""
    
    ws.merge_cells('A1:L1')
    ws['A1'] = "GAP ANALYSIS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:L2')
    ws['A2'] = "Consolidated protection and retention gaps"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Gap Category", 25),
        ("C", "Description", 50),
        ("D", "Affected Systems", 30),
        ("E", "Policy Requirement", 30),
        ("F", "Risk Level", 15),
        ("G", "Remediation Action", 50),
        ("H", "Owner", 25),
        ("I", "Target Date", 15),
        ("J", "Budget Required", 15),
        ("K", "Status", 15),
        ("L", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "GAP-001", "Integrity Protection", "WORM storage not implemented for critical logs",
        "10 critical systems", "ISMS-POL-A.8.15-S2.2.4", "High",
        "Implement cloud object lock for critical log storage", "IT Ops",
        "31.03.2026", "Yes", "Open", "Budget approval pending"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows
    for data_row in range(9, 101):
        # Column A: Auto-generate Gap ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","GAP-"&TEXT(ROW()-8,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        
        # Other columns: Input
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column I: Target Date
        apply_style(ws[f'I{data_row}'], styles['input_cell'])
        ws[f'I{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    category_dv = DataValidation(type="list",
        formula1='"Access Control,Integrity Protection,Transmission Security,Retention Non-Compliance,Backup,Disposal,Separation of Duties,Other"',
        allow_blank=True)
    ws.add_data_validation(category_dv)
    category_dv.add('B9:B100')
    
    risk_dv = DataValidation(type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add('F9:F100')
    
    budget_dv = DataValidation(type="list",
        formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(budget_dv)
    budget_dv.add('J9:J100')
    
    status_dv = DataValidation(type="list",
        formula1='"\u274C Open,⏳ In Progress,\u2705 Resolved,⭕ Deferred"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('K9:K100')
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 12.5: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create Evidence Register for audit documentation."""

    ws.merge_cells('A1:L1')
    ws['A1'] = "EVIDENCE REGISTER"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:L2')
    ws['A2'] = "Document all supporting evidence for log protection & retention assessment"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A3:L3')
    ws['A3'] = "Reference evidence by ID in other assessment sheets. Store evidence files in designated secure location."
    ws['A3'].font = Font(italic=True, size=9)
    ws['A3'].alignment = Alignment(horizontal='left')

    # Evidence table headers
    row = 5
    headers = [
        ("Evidence ID", 12),
        ("Assessment Sheet", 24),
        ("Evidence Type", 20),
        ("Evidence Title", 35),
        ("Description", 40),
        ("File Location/Link", 35),
        ("Date Collected", 14),
        ("Collected By", 18),
        ("Retention Period", 15),
        ("Review Date", 14),
        ("Status", 12),
        ("Notes", 30),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles['column_header'])
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    row += 1

    # Pre-populate with common evidence types for this assessment
    evidence_items = [
        ("EVD-001", "Access Control Assessment", "Configuration", "RBAC Policy Export", "", "", "", "", "7 years", "", "", ""),
        ("EVD-002", "Integrity Protection", "Configuration", "WORM Storage Config", "", "", "", "", "7 years", "", "", ""),
        ("EVD-003", "Integrity Protection", "Report", "Hash Verification Report", "", "", "", "", "7 years", "", "", ""),
        ("EVD-004", "Secure Transmission", "Certificate", "TLS Certificate Details", "", "", "", "", "7 years", "", "", ""),
        ("EVD-005", "Retention Period Compliance", "Policy", "Retention Policy Document", "", "", "", "", "Permanent", "", "", ""),
        ("EVD-006", "Disposal Procedures", "Procedure", "Secure Deletion SOP", "", "", "", "", "Permanent", "", "", ""),
        ("EVD-007", "Legal Hold Management", "Log", "Active Legal Holds List", "", "", "", "", "Per hold", "", "", ""),
        ("EVD-008", "Privacy Impact Assessment", "Assessment", "DPIA Document", "", "", "", "", "7 years", "", "", ""),
    ]

    for evidence_data in evidence_items:
        for col_idx, value in enumerate(evidence_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 4:
                apply_style(cell, styles['input_cell'])
        row += 1

    # Add 50 more empty rows
    for i in range(50):
        evd_num = f"EVD-{9+i:03d}"
        ws.cell(row=row, column=1, value=evd_num)
        for col_idx in range(2, 13):
            cell = ws.cell(row=row, column=col_idx, value="")
            apply_style(cell, styles['input_cell'])
        row += 1

    # Data validations
    type_dv = DataValidation(type="list",
        formula1='"Configuration,Screenshot,Export,Report,Certificate,Policy,Procedure,Log,Assessment,Test Result,Approval,Other"',
        allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add('C6:C100')

    sheet_dv = DataValidation(type="list",
        formula1='"Access Control Assessment,Integrity Protection,Secure Transmission,Retention Period Compliance,Storage Tier Implementation,Log Backup & Recovery,Disposal Procedures,Separation of Duties,Legal Hold Management,Privacy Impact Assessment,Gap Analysis"',
        allow_blank=True)
    ws.add_data_validation(sheet_dv)
    sheet_dv.add('B6:B100')

    status_dv = DataValidation(type="list",
        formula1='"Collected,Pending,Expired,Not Available"',
        allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('K6:K100')

    ws.freeze_panes = 'A6'


# ============================================================================
# SECTION 13: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard sheet."""
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "LOG PROTECTION & RETENTION DASHBOARD"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "Executive Summary - Protection & Compliance Status"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 25
    
    # Section 1: Protection Compliance Summary
    row = 4
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "1. PROTECTION COMPLIANCE SUMMARY"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    protection_metrics = [
        ("Access Controls Compliant", "=COUNTIF('Access Control Assessment'!N:N,\">0.95\")", ">95%"),
        ("Integrity Protection Implemented", "=COUNTIF('Integrity Protection'!N:N,\"Compliant\")", ">95%"),
        ("Secure Transmission (Critical)", "=COUNTIF('Secure Transmission'!K:K,\"Compliant\")", "100%"),
        ("Backup Implemented", "=COUNTIF('Log Backup & Recovery'!O:O,\"Compliant\")", ">95%"),
        ("Separation of Duties", "=COUNTIF('Separation of Duties'!L:L,\"Compliant\")", ">90%"),
        ("Secure Disposal Procedures", "=COUNTIF('Disposal Procedures'!K:K,\"Compliant\")", "100%"),
    ]
    
    ws[f'A{row}'] = "Protection Measure"
    ws[f'B{row}'] = "Compliant Count"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    row += 1
    for measure, formula, target in protection_metrics:
        ws[f'A{row}'] = measure
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles['formula_cell'])
        ws[f'C{row}'] = target
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{row}'] = "=IF(B{row}>0,\"✓\",\"✗\")"
        apply_style(ws[f'D{row}'], styles['formula_cell'])
        row += 1
    
    # Section 2: Retention Compliance Summary
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "2. RETENTION COMPLIANCE SUMMARY"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    retention_headers = ["Log Category", "Compliant", "Under-Retained", "Over-Retained", "% Compliant"]
    for col_idx, header in enumerate(retention_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    categories = ["Security", "Authentication", "Administrative", "Application"]
    
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = f'=COUNTIFS(\'Retention Period Compliance\'!B:B,"{category}",\'Retention Period Compliance\'!O:O,"Compliant")'
        ws[f'C{row}'] = f'=COUNTIFS(\'Retention Period Compliance\'!B:B,"{category}",\'Retention Period Compliance\'!J:J,">0")'
        ws[f'D{row}'] = f'=COUNTIFS(\'Retention Period Compliance\'!B:B,"{category}",\'Retention Period Compliance\'!K:K,">0")'
        ws[f'E{row}'] = f'=IF((B{row}+C{row}+D{row})=0,0,B{row}/(B{row}+C{row}+D{row}))'
        ws[f'E{row}'].number_format = '0.0%'
        
        for col in ['B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles['formula_cell'])
        
        row += 1
    
    # Total row
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'] = f'=SUM({col}{row-4}:{col}{row-1})'
        apply_style(ws[f'{col}{row}'], styles['formula_cell'])
        ws[f'{col}{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f'=IF((B{row}+C{row}+D{row})=0,0,B{row}/(B{row}+C{row}+D{row}))'
    ws[f'E{row}'].number_format = '0.0%'
    apply_style(ws[f'E{row}'], styles['formula_cell'])
    ws[f'E{row}'].font = Font(bold=True)
    
    # Section 3: Gap Summary
    row += 3
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "3. GAP SUMMARY BY RISK LEVEL"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    gap_headers = ["Risk Level", "Open", "In Progress", "Resolved", "Overdue"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    risk_levels = ["Critical", "High", "Medium", "Low"]
    
    for risk in risk_levels:
        ws[f'A{row}'] = risk
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = f'=COUNTIFS(\'Gap Analysis\'!F:F,"{risk}",\'Gap Analysis\'!K:K,"Open")'
        ws[f'C{row}'] = f'=COUNTIFS(\'Gap Analysis\'!F:F,"{risk}",\'Gap Analysis\'!K:K,"In Progress")'
        ws[f'D{row}'] = f'=COUNTIFS(\'Gap Analysis\'!F:F,"{risk}",\'Gap Analysis\'!K:K,"Resolved")'
        ws[f'E{row}'] = f'=SUMPRODUCT((\'Gap Analysis\'!F:F="{risk}")*((\'Gap Analysis\'!K:K="Open")+(\'Gap Analysis\'!K:K="In Progress"))*(\'Gap Analysis\'!I:I<TODAY())*(\'Gap Analysis\'!I:I<>""))'
        
        for col in ['B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles['formula_cell'])
        
        row += 1
    
    # Section 4: Key Findings
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "4. KEY FINDINGS & RECOMMENDATIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    findings = [
        "CRITICAL ACTIONS:",
        "\u2022 Address all Critical risk gaps within 30 days",
        "\u2022 Implement WORM storage for critical log sources",
        "\u2022 Enable TLS 1.2+ for all log transmission over untrusted networks",
        "",
        "HIGH PRIORITY:",
        "\u2022 Implement separation of duties where system admins manage logs",
        "\u2022 Enable integrity protection (hashing) for high-criticality logs",
        "\u2022 Ensure all backups are encrypted and tested quarterly",
        "",
        "RETENTION COMPLIANCE:",
        "\u2022 Address under-retention issues to meet regulatory requirements",
        "\u2022 Implement automated disposal with legal hold checking",
        "\u2022 Review over-retention to optimize storage costs",
        "",
        "CONTINUOUS IMPROVEMENT:",
        "\u2022 Quarterly access control reviews",
        "\u2022 Semi-annual retention period reviews",
        "\u2022 Annual backup restore testing",
    ]
    
    for finding in findings:
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = finding
        if finding.endswith(":"):
            ws[f'A{row}'].font = Font(bold=True, size=10)
        else:
            ws[f'A{row}'].font = Font(size=9)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row += 1
    
    set_column_widths(ws, {
        'A': 35, 'B': 20, 'C': 15, 'D': 15,
        'E': 20, 'F': 20, 'G': 20, 'H': 20
    })


# ============================================================================
# SECTION 14: APPROVAL & SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff_sheet(ws, styles):
    """Create Approval & Sign-Off sheet."""
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "APPROVAL & SIGN-OFF"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "Multi-level approval workflow for Log Protection & Retention Assessment"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 25
    
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "APPROVAL WORKFLOW"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    approval_headers = ["Role", "Name", "Date", "Signature", "Status"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    approval_roles = [
        ("Log Administrator", "[Name]", "", "_____", "☐ Reviewed"),
        ("IT Operations Manager", "[Name]", "", "_____", "☐ Reviewed"),
        ("Information Security Manager", "[Name]", "", "_____", "☐ Approved"),
        ("Legal/Compliance Officer", "[Name]", "", "_____", "☐ Reviewed"),
    ]
    
    for role, name, date, signature, status in approval_roles:
        ws[f'A{row}'] = role
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = name
        apply_style(ws[f'B{row}'], styles['input_cell'])
        ws[f'C{row}'] = date
        apply_style(ws[f'C{row}'], styles['input_cell'])
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'
        ws[f'D{row}'] = signature
        apply_style(ws[f'D{row}'], styles['input_cell'])
        ws[f'E{row}'] = status
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ACKNOWLEDGMENTS CHECKLIST"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    acknowledgments = [
        "☐ Access controls assessed for all log sources",
        "☐ Integrity protection mechanisms evaluated",
        "☐ Secure transmission verified",
        "☐ Retention periods validated against policy",
        "☐ Storage tiers reviewed",
        "☐ Backup and recovery capabilities tested",
        "☐ Disposal procedures assessed",
        "☐ Separation of duties verified",
        "☐ Legal holds documented",
        "☐ Gaps identified and remediation planned",
    ]
    
    for acknowledgment in acknowledgments:
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = acknowledgment
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    
    set_column_widths(ws, {'A': 35, 'B': 25, 'C': 15, 'D': 15, 'E': 20})


# ============================================================================
# SECTION 15: CONDITIONAL FORMATTING & MAIN EXECUTION
# ============================================================================

def apply_conditional_formatting(wb):
    """Apply conditional formatting across sheets."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Access Control Assessment - Compliance Score
    ws = wb['Access Control Assessment']
    ws.conditional_formatting.add('N9:N100',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.95'], fill=green_fill))
    ws.conditional_formatting.add('N9:N100',
        CellIsRule(operator='between', formula=['0.8', '0.94'], fill=yellow_fill))
    ws.conditional_formatting.add('N9:N100',
        CellIsRule(operator='lessThan', formula=['0.8'], fill=red_fill))
    
    # Retention Period Compliance - Compliance Status
    ws = wb['Retention Period Compliance']
    ws.conditional_formatting.add('O9:O100',
        CellIsRule(operator='equal', formula=['"Compliant"'], fill=green_fill))
    ws.conditional_formatting.add('O9:O100',
        CellIsRule(operator='equal', formula=['"Non-Compliant"'], fill=red_fill))
    
    # Gap Analysis - Risk Level
    ws = wb['Gap Analysis']
    ws.conditional_formatting.add('F9:F100',
        CellIsRule(operator='equal', formula=['"Critical"'], fill=red_fill))
    ws.conditional_formatting.add('F9:F100',
        CellIsRule(operator='equal', formula=['"High"'], fill=orange_fill))
    ws.conditional_formatting.add('F9:F100',
        CellIsRule(operator='equal', formula=['"Medium"'], fill=yellow_fill))
    ws.conditional_formatting.add('F9:F100',
        CellIsRule(operator='equal', formula=['"Low"'], fill=green_fill))
    
    # Gap Analysis - Status
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"Open"'], fill=red_fill))
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"In Progress"'], fill=yellow_fill))
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"Resolved"'], fill=green_fill))


def main():
    """
    Main execution function.
    
    "Security is always excessive until it's not enough." - Robbie Sinclair
    Let's make sure it's just right!
    """
    
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.15.3 - Log Protection & Retention Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.15: Logging")
    logger.info("=" * 78)
    logger.info("")
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("[1/15] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    logger.info("[2/15] Creating Access Control Assessment...")
    create_access_control_sheet(wb["Access Control Assessment"], styles)

    logger.info("[3/15] Creating Integrity Protection...")
    create_integrity_protection_sheet(wb["Integrity Protection"], styles)

    logger.info("[4/15] Creating Secure Transmission...")
    create_secure_transmission_sheet(wb["Secure Transmission"], styles)

    logger.info("[5/15] Creating Retention Period Compliance...")
    create_retention_period_sheet(wb["Retention Period Compliance"], styles)

    logger.info("[6/15] Creating Storage Tier Implementation...")
    create_storage_tier_sheet(wb["Storage Tier Implementation"], styles)

    logger.info("[7/15] Creating Log Backup & Recovery...")
    create_backup_recovery_sheet(wb["Log Backup & Recovery"], styles)

    logger.info("[8/15] Creating Disposal Procedures...")
    create_disposal_procedures_sheet(wb["Disposal Procedures"], styles)

    logger.info("[9/15] Creating Separation of Duties...")
    create_separation_of_duties_sheet(wb["Separation of Duties"], styles)

    logger.info("[10/15] Creating Legal Hold Management...")
    create_legal_hold_sheet(wb["Legal Hold Management"], styles)

    logger.info("[11/15] Creating Privacy Impact Assessment...")
    create_privacy_impact_sheet(wb["Privacy Impact Assessment"], styles)

    logger.info("[12/15] Creating Evidence Register...")
    create_evidence_register_sheet(wb["Evidence Register"], styles)

    logger.info("[13/15] Creating Gap Analysis...")
    create_gap_analysis_sheet(wb["Gap Analysis"], styles)

    logger.info("[14/15] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[15/15] Creating Approval & Sign-Off...")
    create_approval_signoff_sheet(wb["Approval & Sign-Off"], styles)
    
    logger.info("")
    logger.info("Applying conditional formatting...")
    apply_conditional_formatting(wb)
    
    filename = OUTPUT_FILENAME
    
    logger.info("")
    logger.info("Saving workbook...")
    wb.save(filename)
    
    logger.info("")
    logger.info("=" * 78)
    logger.info("\u2705 SUCCESS: Workbook generated successfully!")
    logger.info("=" * 78)
    logger.info("")
    logger.info(f"📄 Filename: {filename}")
    logger.info(f"📊 Estimated file size: ~650 KB - 1.1 MB")
    logger.info("")
    logger.info("Workbook Structure:")
    logger.info("  Y Sheet 1:  Instructions & Legend")
    logger.info("  Y Sheet 2:  Access Control Assessment (92 rows)")
    logger.info("  Y Sheet 3:  Integrity Protection (92 rows)")
    logger.info("  Y Sheet 4:  Secure Transmission (192 rows)")
    logger.info("  Y Sheet 5:  Retention Period Compliance (92 rows)")
    logger.info("  Y Sheet 6:  Storage Tier Implementation (22 rows)")
    logger.info("  Y Sheet 7:  Log Backup & Recovery (17 rows)")
    logger.info("  Y Sheet 8:  Disposal Procedures (42 rows)")
    logger.info("  Y Sheet 9:  Separation of Duties (42 rows)")
    logger.info("  Y Sheet 10: Legal Hold Management (22 rows)")
    logger.info("  Y Sheet 11: Privacy Impact Assessment (42 rows)")
    logger.info("  Y Sheet 12: Evidence Register (22 rows)")
    logger.info("  Y Sheet 13: Gap Analysis (92 rows)")
    logger.info("  Y Sheet 14: Summary Dashboard (with compliance metrics)")
    logger.info("  Y Sheet 15: Approval & Sign-Off")
    logger.info("")
    logger.info("Features:")
    logger.info("  Y Access control scoring")
    logger.info("  Y Integrity protection assessment")
    logger.info("  Y Retention compliance calculations")
    logger.info("  Y Auto-generated IDs (Hold, Gap, Evidence)")
    logger.info("  Y Conditional formatting (Green/Yellow/Red)")
    logger.info("  Y Legal hold tracking")
    logger.info("  Y Privacy impact assessment (GDPR/nFADP)")
    logger.info("  Y Evidence register with audit trail")
    logger.info("  Y Date format: DD.MM.YYYY")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Assess access controls for all log sources")
    logger.info("  2. Evaluate integrity protection mechanisms")
    logger.info("  3. Verify secure transmission configurations")
    logger.info("  4. Validate retention periods against policy")
    logger.info("  5. Review storage tier implementation")
    logger.info("  6. Test backup and recovery procedures")
    logger.info("  7. Verify disposal procedures compliance")
    logger.info("  8. Assess separation of duties")
    logger.info("  9. Document any active legal holds")
    logger.info(" 10. Complete privacy impact assessment")
    logger.info(" 11. Maintain evidence register for audit")
    logger.info(" 12. Create remediation plans for identified gaps")
    logger.info("")
    logger.info("═" * 78)
    logger.info("'Trust, but verify. And get it in writing.' - Security Proverb")
    logger.info("Document your log protection properly!")
    logger.info("═" * 78)
    logger.info("")


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
