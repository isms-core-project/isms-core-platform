#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.15.1 - Log Source Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.15: Logging
Assessment Domain 1 of 4: System Inventory and Log Source Cataloging

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific logging infrastructure, system architecture, and
assessment requirements.

Key customisation areas:
1. System types and categories (match your infrastructure taxonomy)
2. Log source priorities and criticality levels (adapt to your risk profile)
3. Regulatory scope mappings (specific to your compliance requirements)
4. Data classification schemes (based on your information security policy)
5. Compliance scoring criteria (aligned with your ISMS maturity targets)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.15 Logging Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for documenting
and evaluating log sources across the organisation's entire IT infrastructure.

**Purpose:**
Enables systematic inventory and assessment of all systems requiring logging
against ISO 27001:2022 Control A.8.15 requirements, supporting evidence-based
validation of logging coverage and completeness.

**Assessment Scope:**
- System inventory (servers, workstations, network devices, applications)
- Log event types by system (authentication, authorisation, administrative)
- Authentication logging capabilities and implementation
- Authorisation and access control logging
- Administrative activity logging
- Security event logging (intrusions, malware, anomalies)
- Application and database logging
- Network device logging (firewalls, routers, switches, IDS/IPS)
- Gap analysis for systems without adequate logging
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and logging standards
2. System Inventory - Complete catalog of in-scope systems
3. Log Event Types by System - Event type mapping per system
4. Authentication Logging - Login/logout event assessment
5. Authorisation & Access - Permission and access control logging
6. Administrative Activity - Privileged operation logging
7. Security Event Logging - Security-relevant event assessment
8. Application & Database - Application-specific logging
9. Network Device Logging - Network infrastructure logging
10. Gap Analysis - Systems lacking adequate logging
11. Evidence Register - Audit evidence tracking and documentation
12. Summary Dashboard - Compliance metrics and KPIs
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dropdown lists for consistency
- Conditional formatting for compliance status visualization
- Automated gap identification for systems without logging
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with SIEM and log management platforms

**Integration:**
This assessment feeds into the A.8.15.5 Compliance Dashboard, which
consolidates data from all four logging assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a815_1_log_source_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a815_1_log_source_inventory.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a815_1_log_source_inventory.py --date 20250124

Output:
    File: ISMS_A_8_15_1_Log_Source_Inventory_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customise system categories to match your environment
    2. Inventory all systems in scope for logging requirements
    3. Complete log source assessments for each system
    4. Validate logging capabilities against requirements
    5. Review gap analysis for systems without adequate logging
    6. Define remediation actions with timelines
    7. Collect and link audit evidence (configurations, screenshots)
    8. Obtain stakeholder approvals
    9. Feed results into A.8.15.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.15
Assessment Domain:    1 of 4 (System Inventory and Log Source Cataloging)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.15: Logging Policy (Governance)
    - ISMS-IMP-A.8.15.1: Log Source Inventory Implementation Guide
    - ISMS-IMP-A.8.15.2: Log Collection & Centralization Assessment (Domain 2)
    - ISMS-IMP-A.8.15.3: Log Protection & Retention Assessment (Domain 3)
    - ISMS-IMP-A.8.15.4: Log Analysis & Review Assessment (Domain 4)
    - ISMS-IMP-A.8.15.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.15.1 specification
    - Supports comprehensive log source inventory and cataloging
    - Integrated with A.8.15.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Logging Standards:**
Logging requirements evolve with threat landscape changes. Review industry
standards (NIST SP 800-92, CIS Critical Security Controls, MITRE ATT&CK)
quarterly and update assessment criteria accordingly. Ensure logging coverage
aligns with current security best practices.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of logging capabilities for all critical
and high-risk systems.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- System inventory and network topology
- Security control implementation details
- Vulnerability information and security gaps
- Regulatory compliance scope

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new systems added to infrastructure
- Semi-annually: Update logging requirements based on threat landscape
- Annually: Complete reassessment of all systems
- Ad-hoc: When infrastructure changes or new compliance requirements emerge

**Quality Assurance:**
Have logging/SIEM SMEs and security operations engineers validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Ensure logging requirements align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 logging and monitoring requirements
- Healthcare: HIPAA audit trail requirements
- Finance: Regional financial services logging requirements
- Government: Jurisdiction-specific audit logging mandates

Customise assessment criteria to include regulatory-specific requirements.

**SIEM Integration:**
This assessment identifies log sources that should be forwarded to SIEM.
Coordinate with SOC team to ensure identified log sources are configured
for centralized collection and analysis.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Document identification constants

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.15.1"
WORKBOOK_NAME = "Log Source Inventory Assessment"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
CONTROL_ID   = "A.8.15"
CONTROL_NAME = "Logging"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

from openpyxl.chart import BarChart, Reference

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.15.1 specification
    sheets = [
        "Instructions & Legend",
        "System Inventory",
        "Log Event Types by System",
        "Authentication Logging",
        "Authorisation & Access",
        "Administrative Activity",
        "Security Event Logging",
        "Application & Database",
        "Network Device Logging",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Returns style TEMPLATES (dictionaries) to avoid shared object warnings.
    Each cell gets its own style instance to prevent Excel repair issues.
    
    Philosophy: "The first principle is that you must not fool yourself"
    - Feynman. Don't cargo cult shared styles - create fresh ones!
    """
    return {
        'header_main': {
            'font': {'name': 'Calibri', 'size': 14, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '003366', 'end_color': '003366', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'header_sub': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '003366', 'end_color': '003366', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'column_header': {
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '000000'},
            'fill': {'start_color': 'D9D9D9', 'end_color': 'D9D9D9', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'input_cell': {
            'fill': {'start_color': 'FFFFCC', 'end_color': 'FFFFCC', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'example_cell': {
            'fill': {'start_color': 'F2F2F2', 'end_color': 'F2F2F2', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'italic': True, 'color': '666666'},
            'alignment': {'horizontal': 'left', 'vertical': 'top'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'formula_cell': {
            'fill': {'start_color': 'F2F2F2', 'end_color': 'F2F2F2', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'status_compliant': {
            'fill': {'start_color': 'C6EFCE', 'end_color': 'C6EFCE', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '006100'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'status_partial': {
            'fill': {'start_color': 'FFEB9C', 'end_color': 'FFEB9C', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '9C6500'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'status_noncompliant': {
            'fill': {'start_color': 'FFC7CE', 'end_color': 'FFC7CE', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': 'C00000'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'info_cell': {
            'fill': {'start_color': 'F2F2F2', 'end_color': 'F2F2F2', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        }
    }



_STYLES = setup_styles()
def apply_style(cell, style_template):
    """
    Apply a style template to a cell.
    Creates NEW objects for each cell to avoid Excel repair warnings.
    """
    if 'font' in style_template:
        cell.font = Font(**style_template['font'])
    if 'fill' in style_template:
        cell.fill = PatternFill(**style_template['fill'])
    if 'alignment' in style_template:
        cell.alignment = Alignment(**style_template['alignment'])
    if 'border' in style_template:
        cell.border = Border(
            left=Side(style=style_template['border'].get('left', 'thin')),
            right=Side(style=style_template['border'].get('right', 'thin')),
            top=Side(style=style_template['border'].get('top', 'thin')),
            bottom=Side(style=style_template['border'].get('bottom', 'thin'))
        )


def set_column_widths(ws, widths):
    """Set column widths from a dictionary {column_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the ‘System Inventory’ sheet for all in-scope systems.', '2. For each system, document log sources and event types logged.', '3. Use dropdowns to ensure consistent data entry.', '4. Yellow cells require user input, grey cells are auto-calculated.', '5. Complete compliance checklist for each log source type.', '6. Review Summary Dashboard for overall compliance status.', '7. Document gaps in Gap Analysis sheet.', '8. Develop remediation plan for identified gaps.', '9. Obtain approvals in Approval & Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_system_inventory_sheet(ws, styles):
    """
    Create System Inventory sheet - the foundation of log source catalog.
    
    "You can't improve what you don't measure, and you can't measure 
    what you don't catalog." - Adapted from Peter Drucker
    """
    
    # Header
    ws.merge_cells('A1:Q1')
    ws['A1'] = "SYSTEM INVENTORY"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells('A2:Q2')
    ws['A2'] = "Document all systems in scope for logging. Include servers, network devices, applications, cloud services, and security tools."
    apply_style(ws['A2'], styles['header_sub'])

    # Column headers (Row 4)
    headers = [
        ("A", "System ID", 15),
        ("B", "System Name", 30),
        ("C", "System Type", 20),
        ("D", "Operating System / Platform", 25),
        ("E", "Environment", 15),
        ("F", "Data Classification", 18),
        ("G", "Business Criticality", 18),
        ("H", "Regulatory Scope", 20),
        ("I", "Logging Priority", 15),
        ("J", "System Owner", 25),
        ("K", "Owner Email", 30),
        ("L", "Hostname / FQDN", 30),
        ("M", "Primary IP", 15),
        ("N", "Location", 20),
        ("O", "Logging Enabled", 15),
        ("P", "Forwarding to SIEM", 18),
        ("Q", "Compliance Status", 18),
    ]

    row = 4
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width

    # Example row (Row 5 - Gray italic)
    row = 5
    example_data = [
        "SYS-001", "web-prod-01", "Server", "Linux", "Production",
        "Confidential", "Critical (T1)", "PCI DSS v4.0.1, GDPR", "P1-Critical",
        "John Doe", "jdoe@example.ch", "web-prod-01.example.com",
        "10.0.1.50", "DC1", "Yes", "Yes", "Compliant"
    ]

    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])

    # Data entry rows (6-55 = 50 rows)
    for data_row in range(6, 56):
        # Column A: Auto-generate System ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","SYS-"&TEXT(ROW()-5,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])

        # Columns B-P: Input cells
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])

        # Column Q: Compliance Status Formula
        ws[f'Q{data_row}'] = f'=IF(B{data_row}="","",IF(AND(O{data_row}="Yes",P{data_row}="Yes"),"Compliant",IF(OR(O{data_row}="Partial",P{data_row}="Planned"),"Partial","Non-Compliant")))'
        apply_style(ws[f'Q{data_row}'], styles['formula_cell'])

    # Data validations
    validations = []

    system_type_dv = DataValidation(type="list",
        formula1='"Server,Network Device,Security Appliance,Application,Cloud Service,Database,Other"',
        allow_blank=False)
    system_type_dv.add('C6:C55')
    validations.append(system_type_dv)

    os_platform_dv = DataValidation(type="list",
        formula1='"Windows,Linux,Unix,Network OS,Cloud,Application Platform,Other"',
        allow_blank=False)
    os_platform_dv.add('D6:D55')
    validations.append(os_platform_dv)

    environment_dv = DataValidation(type="list",
        formula1='"Production,Staging,Development,Test"',
        allow_blank=False)
    environment_dv.add('E6:E55')
    validations.append(environment_dv)

    data_class_dv = DataValidation(type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=False)
    data_class_dv.add('F6:F55')
    validations.append(data_class_dv)

    criticality_dv = DataValidation(type="list",
        formula1='"Critical (T1),High (T2),Medium (T3),Low (T4)"',
        allow_blank=False)
    criticality_dv.add('G6:G55')
    validations.append(criticality_dv)

    priority_dv = DataValidation(type="list",
        formula1='"P1-Critical,P2-High,P3-Medium,P4-Low"',
        allow_blank=False)
    priority_dv.add('I6:I55')
    validations.append(priority_dv)

    logging_enabled_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,\u26A0\uFE0F Partial,\u2753 Unknown"',
        allow_blank=False)
    logging_enabled_dv.add('O6:O55')
    validations.append(logging_enabled_dv)

    forwarding_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,\u231B Planned,\u2796 N/A"',
        allow_blank=False)
    forwarding_dv.add('P6:P55')
    validations.append(forwarding_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A6'

# ============================================================================
# SECTION 4: LOG EVENT TYPES BY SYSTEM SHEET
# ============================================================================

def create_log_event_types_sheet(ws, styles):
    """
    Create Log Event Types sheet - documents what each system logs.
    
    "The devil is in the details, but so is salvation." - Hyman Rickover
    Let's get granular about what's actually being logged!
    """
    
    # Header
    ws.merge_cells('A1:S1')
    ws['A1'] = "LOG EVENT TYPES BY SYSTEM"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells('A2:S2')
    ws['A2'] = "For each system, document which event types are logged. This assesses logging completeness per ISMS-POL-A.8.15-S2.1"
    apply_style(ws['A2'], styles['header_sub'])

    # Column headers (Row 4)
    headers = [
        ("A", "System ID", 15),
        ("B", "System Name", 30),
        ("C", "Authentication Events", 15),
        ("D", "Authorisation Events", 15),
        ("E", "Administrative Actions", 15),
        ("F", "Security Events", 15),
        ("G", "Application Events", 15),
        ("H", "System Events", 15),
        ("I", "Network Events", 15),
        ("J", "Database Events", 15),
        ("K", "Log Format", 15),
        ("L", "Timestamp Format", 18),
        ("M", "Timezone", 15),
        ("N", "Est. Daily Volume (MB)", 20),
        ("O", "Retention Period (months)", 20),
        ("P", "Storage Tier", 15),
        ("Q", "Protection Mechanisms", 20),
        ("R", "Event Types Completeness", 20),
        ("S", "Notes", 40),
    ]

    row = 4
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width

    # Example row (Row 5)
    row = 5
    example_data = [
        "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Yes",
        "Partial", "Partial", "N/A", "Syslog", "ISO 8601", "UTC",
        "150", "12", "Hot", "Access Controls, Encryption", "87.5%",
        "Web server logs include app events"
    ]

    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])

    # Data entry rows (6-55 = 50 rows)
    for data_row in range(6, 56):
        # Column A: System ID dropdown (links to System Inventory)
        apply_style(ws[f'A{data_row}'], styles['input_cell'])

        # Column B: VLOOKUP System Name
        ws[f'B{data_row}'] = f'=IF(A{data_row}="","",IFERROR(VLOOKUP(A{data_row},\'System Inventory\'!A:B,2,FALSE),"Not Found"))'
        apply_style(ws[f'B{data_row}'], styles['formula_cell'])

        # Columns C-J: Event type dropdowns (Yes/No/Partial/N/A)
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])

        # Columns K-M: Format dropdowns
        for col_letter in ['K', 'L', 'M']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])

        # Columns N-O: Numeric inputs
        for col_letter in ['N', 'O']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])

        # Column P: Storage tier dropdown
        apply_style(ws[f'P{data_row}'], styles['input_cell'])

        # Column Q: Protection mechanisms (text)
        apply_style(ws[f'Q{data_row}'], styles['input_cell'])

        # Column R: Event Types Completeness Formula
        ws[f'R{data_row}'] = f'=IF(A{data_row}="","",COUNTIF(C{data_row}:J{data_row},"*Yes*")/COUNTA(C{data_row}:J{data_row}))'
        ws[f'R{data_row}'].number_format = '0.0%'
        apply_style(ws[f'R{data_row}'], styles['formula_cell'])

        # Column S: Notes
        apply_style(ws[f'S{data_row}'], styles['input_cell'])

    # Data validations
    validations = []

    yes_no_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,\u26A0\uFE0F Partial,\u2796 N/A"',
        allow_blank=True)
    yes_no_dv.add('C6:J55')
    validations.append(yes_no_dv)

    log_format_dv = DataValidation(type="list",
        formula1='"Syslog,CEF,JSON,EVTX,Custom,Unknown"',
        allow_blank=True)
    log_format_dv.add('K6:K55')
    validations.append(log_format_dv)

    timestamp_format_dv = DataValidation(type="list",
        formula1='"ISO 8601,RFC 3339,Unix Epoch,Custom,Unknown"',
        allow_blank=True)
    timestamp_format_dv.add('L6:L55')
    validations.append(timestamp_format_dv)

    timezone_dv = DataValidation(type="list",
        formula1='"UTC,Local,Unknown"',
        allow_blank=True)
    timezone_dv.add('M6:M55')
    validations.append(timezone_dv)

    storage_tier_dv = DataValidation(type="list",
        formula1='"Hot,Warm,Cold,Unknown"',
        allow_blank=True)
    storage_tier_dv.add('P6:P55')
    validations.append(storage_tier_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'C5'

# ============================================================================
# SECTION 5: ASSESSMENT SHEET TEMPLATE FUNCTION
# ============================================================================

def create_assessment_sheet(ws, styles, sheet_config):
    """
    Generic function to create assessment sheets (Authentication, Authorisation, etc.)
    
    sheet_config = {
        'title': 'Sheet title',
        'description': 'Instructions text',
        'assessment_columns': [list of tuples (col_letter, header, width)]
    }
    
    "Don't repeat yourself" - DRY principle. One function, many sheets!
    """
    
    # Header
    end_col = sheet_config['assessment_columns'][-1][0]
    ws.merge_cells(f'A1:{end_col}1')
    ws['A1'] = sheet_config['title']
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells(f'A2:{end_col}2')
    ws['A2'] = sheet_config['description']
    apply_style(ws['A2'], styles['header_sub'])

    # Column headers (Row 4)
    row = 4
    for col_letter, header, width in sheet_config['assessment_columns']:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width

    # Example row (Row 5)
    row = 5
    if 'example_data' in sheet_config:
        for col_idx, value in enumerate(sheet_config['example_data'], start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            apply_style(cell, styles['example_cell'])

    # Data entry rows (6-55 = 50 rows)
    for data_row in range(6, 56):
        # Column A: System ID dropdown
        apply_style(ws[f'A{data_row}'], styles['input_cell'])

        # Column B: VLOOKUP System Name
        ws[f'B{data_row}'] = f'=IF(A{data_row}="","",IFERROR(VLOOKUP(A{data_row},\'System Inventory\'!A:B,2,FALSE),"Not Found"))'
        apply_style(ws[f'B{data_row}'], styles['formula_cell'])

        # Columns C through second-to-last: Input cells
        for col_letter, header, width in sheet_config['assessment_columns'][2:-3]:  # Skip A, B and last 3 cols
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])

        # Column P (or configured): Compliance Score Formula
        score_col = sheet_config.get('score_column', 'P')
        # Get the range of assessment columns (C through O typically)
        first_assess_col = sheet_config['assessment_columns'][2][0]
        last_assess_col = chr(ord(score_col) - 3)  # 3 columns before score column

        ws[f'{score_col}{data_row}'] = f'=IF(A{data_row}="","",(COUNTIF({first_assess_col}{data_row}:{last_assess_col}{data_row},"*Yes*")/COUNTIF({first_assess_col}{data_row}:{last_assess_col}{data_row},"<>N/A"))*100)'
        ws[f'{score_col}{data_row}'].number_format = '0.0"%"'
        apply_style(ws[f'{score_col}{data_row}'], styles['formula_cell'])

        # Last two columns: Gap Description and Remediation Plan (text input)
        gap_col = chr(ord(score_col) + 1)
        remediation_col = chr(ord(score_col) + 2)
        apply_style(ws[f'{gap_col}{data_row}'], styles['input_cell'])
        apply_style(ws[f'{remediation_col}{data_row}'], styles['input_cell'])

    # Add data validations for Yes/No/Partial/N/A columns
    validations = []
    yes_no_cols = [col[0] for col in sheet_config['assessment_columns'][2:-3]]
    for col_letter in yes_no_cols:
        yes_no_dv = DataValidation(type="list",
            formula1='"\u2705 Yes,\u274C No,\u26A0\uFE0F Partial,\u2796 N/A"',
            allow_blank=True)
        yes_no_dv.add(f'{col_letter}6:{col_letter}55')
        validations.append(yes_no_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'C5'

# ============================================================================
# SECTION 6: AUTHENTICATION LOGGING ASSESSMENT SHEET
# ============================================================================

def create_authentication_logging_sheet(ws, styles):
    """
    Create Authentication Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.2 - Authentication events.
    """
    
    config = {
        'title': 'AUTHENTICATION LOGGING ASSESSMENT',
        'description': 'Assess authentication event logging completeness per ISMS-POL-A.8.15-S2.1.2',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.2 (Authentication Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Logs Successful Logins", 18),
            ("D", "Logs Failed Logins", 18),
            ("E", "Logs Account Lockouts", 18),
            ("F", "Logs Password Changes", 18),
            ("G", "Logs Session Start/End", 18),
            ("H", "Includes User ID", 15),
            ("I", "Includes Timestamp", 15),
            ("J", "Includes Source IP", 15),
            ("K", "Includes Auth Method", 15),
            ("L", "MFA Events Logged", 15),
            ("M", "SSO Events Logged", 15),
            ("N", "Service Account Auth", 20),
            ("O", "Privileged Auth Logged", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Yes",
            "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "N/A", "Yes",
            "100.0%", "", ""
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 7: AUTHORIZATION & ACCESS LOGGING SHEET
# ============================================================================

def create_authorisation_logging_sheet(ws, styles):
    """
    Create Authorisation & Access Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.3 - Authorisation and access control events.
    """
    
    config = {
        'title': 'AUTHORIZATION & ACCESS LOGGING ASSESSMENT',
        'description': 'Assess access control event logging per ISMS-POL-A.8.15-S2.1.3',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.3 (Authorisation Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Logs Access Grants", 18),
            ("D", "Logs Access Denials", 18),
            ("E", "Logs Permission Changes", 18),
            ("F", "Logs Privilege Escalation", 18),
            ("G", "Logs Data Access (Sensitive)", 18),
            ("H", "Includes User ID", 15),
            ("I", "Includes Resource Accessed", 15),
            ("J", "Includes Action Type", 15),
            ("K", "Includes Outcome", 15),
            ("L", "Includes Reason (denied)", 15),
            ("M", "Includes Before/After State", 15),
            ("N", "Logs Group Membership", 20),
            ("O", "Logs Role Assignment", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Yes",
            "Yes", "Yes", "Yes", "Yes", "Partial", "Partial", "Yes", "Yes",
            "92.3%", "Before/After state partially logged", "Enhance logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 8: ADMINISTRATIVE ACTIVITY LOGGING SHEET
# ============================================================================

def create_administrative_activity_sheet(ws, styles):
    """
    Create Administrative Activity Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.4 - Administrative actions.
    """
    
    config = {
        'title': 'ADMINISTRATIVE ACTIVITY LOGGING ASSESSMENT',
        'description': 'Assess administrative action logging per ISMS-POL-A.8.15-S2.1.4',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.4 (Administrative Actions)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "User Account Management", 18),
            ("D", "Group/Role Management", 18),
            ("E", "Configuration Changes", 18),
            ("F", "Security Policy Changes", 18),
            ("G", "Software Install/Uninstall", 18),
            ("H", "Service Start/Stop", 15),
            ("I", "Patch Application", 15),
            ("J", "Bulk Data Operations", 15),
            ("K", "Privileged Session Logging", 15),
            ("L", "Includes Administrator ID", 15),
            ("M", "Includes Before/After Values", 15),
            ("N", "Includes Timestamp", 20),
            ("O", "Includes Change Reason", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Partial",
            "Yes", "Partial", "No", "Yes", "Yes", "Partial", "Yes", "No",
            "76.9%", "Change reason not logged, bulk ops not logged", "Implement change logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 9: SECURITY EVENT LOGGING SHEET
# ============================================================================

def create_security_event_logging_sheet(ws, styles):
    """
    Create Security Event Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.5 - Security tool and event logging.
    """
    
    config = {
        'title': 'SECURITY EVENT LOGGING ASSESSMENT',
        'description': 'Assess security tool and event logging per ISMS-POL-A.8.15-S2.1.5',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.5 (Security Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Firewall Events", 18),
            ("D", "IDS/IPS Alerts", 18),
            ("E", "Anti-malware Events", 18),
            ("F", "DLP Events", 18),
            ("G", "Web Filtering Events", 18),
            ("H", "Email Gateway Events", 15),
            ("I", "EDR Events", 15),
            ("J", "Vulnerability Scan Results", 15),
            ("K", "Security Incident Events", 15),
            ("L", "Includes Severity Level", 15),
            ("M", "Includes Threat Indicators", 15),
            ("N", "Includes Response Actions", 20),
            ("O", "Automated Response Logged", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "FW-001", "perimeter-fw-01", "Yes", "Yes", "N/A", "N/A", "N/A",
            "N/A", "N/A", "N/A", "Yes", "Yes", "Yes", "Partial", "No",
            "100.0%", "Automated response not logged", "Enable response logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)


# ============================================================================
# SECTION 10: APPLICATION & DATABASE LOGGING SHEET
# ============================================================================

def create_application_database_logging_sheet(ws, styles):
    """
    Create Application & Database Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.7 - Application and database events.
    """
    
    config = {
        'title': 'APPLICATION & DATABASE LOGGING ASSESSMENT',
        'description': 'Assess application and database logging per ISMS-POL-A.8.15-S2.1.7',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.7 (Application & Database Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Web App Access Logging", 18),
            ("D", "API Call Logging", 18),
            ("E", "Transaction Logging", 18),
            ("F", "Application Errors", 18),
            ("G", "Database Connections", 18),
            ("H", "Database Queries (Sensitive)", 15),
            ("I", "Schema Changes (DDL)", 15),
            ("J", "Permission Grants", 15),
            ("K", "Backup/Restore Operations", 15),
            ("L", "Includes User/App Identity", 15),
            ("M", "Includes Data Modified", 15),
            ("N", "Includes Query Text", 20),
            ("O", "Includes Row Count", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "DB-001", "prod-db-01", "N/A", "Yes", "Yes", "Yes", "Yes",
            "Yes", "Yes", "Yes", "Yes", "Yes", "Partial", "Partial", "Yes",
            "91.7%", "Query text truncated, data modified partially logged", "Enable full query logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)


# ============================================================================
# SECTION 11: NETWORK DEVICE LOGGING SHEET
# ============================================================================

def create_network_device_logging_sheet(ws, styles):
    """
    Create Network Device Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.8 - Network infrastructure logging.
    """
    
    config = {
        'title': 'NETWORK DEVICE LOGGING ASSESSMENT',
        'description': 'Assess network infrastructure logging per ISMS-POL-A.8.15-S2.1.8',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.8 (Network Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "Device ID", 15),
            ("B", "Device Name", 30),
            ("C", "Connection Logging", 18),
            ("D", "Rule Match Logging", 18),
            ("E", "Configuration Changes", 18),
            ("F", "Interface Up/Down Events", 18),
            ("G", "Routing Changes", 18),
            ("H", "VPN Session Logging", 15),
            ("I", "DHCP/DNS Events", 15),
            ("J", "Wireless Events", 15),
            ("K", "NAT Translations", 15),
            ("L", "Includes Source/Dest IP", 15),
            ("M", "Includes Ports/Protocols", 15),
            ("N", "Includes Action (allow/deny)", 20),
            ("O", "Includes Bytes Transferred", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "NET-001", "core-switch-01", "Yes", "Yes", "Yes", "Yes", "Yes",
            "N/A", "Yes", "N/A", "N/A", "Yes", "Yes", "Yes", "No",
            "100.0%", "Bytes transferred not logged", "Enable traffic volume logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 12: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """
    Create Gap Analysis sheet - consolidated gaps from all assessments.
    
    "The gap between where you are and where you want to be is called a plan."
    - Unknown (probably an auditor)
    """
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "GAP ANALYSIS - CONSOLIDATED FINDINGS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells('A2:L2')
    ws['A2'] = "Document all identified gaps from assessment sheets. Prioritise by risk and establish remediation plans with clear ownership and target dates."
    apply_style(ws['A2'], styles['header_sub'])

    # Column headers (Row 4)
    headers = [
        ("A", "Gap ID", 12),
        ("B", "System ID", 15),
        ("C", "System Name", 30),
        ("D", "Gap Category", 25),
        ("E", "Gap Description", 50),
        ("F", "Policy Requirement", 30),
        ("G", "Impact / Risk", 20),
        ("H", "Remediation Action", 50),
        ("I", "Responsible Party", 25),
        ("J", "Target Date", 15),
        ("K", "Status", 15),
        ("L", "Notes", 40),
    ]

    row = 4
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width

    # Example row (Row 5)
    row = 5
    example_data = [
        "GAP-001", "SYS-042", "mail-server-01", "Event Type Not Logged",
        "Failed authentication attempts not logged",
        "ISMS-POL-A.8.15-S2.1.2", "High",
        "Configure mail server to log authentication failures",
        "System Admin Team", "31.03.2026", "Open",
        "Vendor support required for config change"
    ]

    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])

    # Data entry rows (6-55 = 50 gap rows)
    _gap_yell = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    for data_row in range(6, 56):
        # Column A: Auto-generate Gap ID (FFFFCC — not F2F2F2 formula_cell)
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","GAP-"&TEXT(ROW()-5,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        ws[f'A{data_row}'].fill = _gap_yell

        # Column B: System ID dropdown (links to System Inventory)
        apply_style(ws[f'B{data_row}'], styles['input_cell'])

        # Column C: VLOOKUP System Name (FFFFCC — not F2F2F2 formula_cell)
        ws[f'C{data_row}'] = f'=IF(B{data_row}="","",IFERROR(VLOOKUP(B{data_row},\'System Inventory\'!A:B,2,FALSE),"Not Found"))'
        apply_style(ws[f'C{data_row}'], styles['formula_cell'])
        ws[f'C{data_row}'].fill = _gap_yell

        # Columns D-I: Input cells
        for col_letter in ['D', 'E', 'F', 'G', 'H', 'I']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])

        # Column J: Target Date (with date format)
        apply_style(ws[f'J{data_row}'], styles['input_cell'])
        ws[f'J{data_row}'].number_format = 'DD.MM.YYYY'

        # Columns K-L: Status and Notes
        for col_letter in ['K', 'L']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])

    # Data validations
    validations = []

    gap_category_dv = DataValidation(type="list",
        formula1='"Log Source Missing,Event Type Not Logged,Incomplete Fields,Format Non-Standard,Protection Inadequate,Other"',
        allow_blank=True)
    gap_category_dv.add('D6:D55')
    validations.append(gap_category_dv)

    risk_dv = DataValidation(type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True)
    risk_dv.add('G6:G55')
    validations.append(risk_dv)

    status_dv = DataValidation(type="list",
        formula1='"\u274C Open,\u231B In Progress,\u2705 Resolved,\u2B55 Deferred"',
        allow_blank=True)
    status_dv.add('K6:K55')
    validations.append(status_dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Summary section below data (Row 60+)
    row = 60
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = "GAP SUMMARY STATISTICS"
    apply_style(ws[f'A{row}'], styles['header_sub'])

    row += 2
    summary_labels = [
        ("Total Gaps Identified:", f'=COUNTA(A6:A55)'),
        ("Open Gaps:", f'=COUNTIF(K6:K55,"*Open*")'),
        ("In Progress:", f'=COUNTIF(K6:K55,"*In Progress*")'),
        ("Resolved:", f'=COUNTIF(K6:K55,"*Resolved*")'),
        ("Critical Priority:", f'=COUNTIF(G6:G55,"*Critical*")'),
        ("High Priority:", f'=COUNTIF(G6:G55,"*High*")'),
        ("Overdue (Past Target Date):", f'=SUMPRODUCT((K6:K55<>"Resolved")*(J6:J55<TODAY())*(J6:J55<>""))'),
    ]

    for label, formula in summary_labels:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles['info_cell'])
        row += 1

    ws.freeze_panes = 'A6'

# ============================================================================
# SECTION 13: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet (golden standard)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header A1:H1
    ws.merge_cells('A1:H1')
    ws['A1'] = "EVIDENCE REGISTER"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle (no banner)
    ws.merge_cells('A2:H2')
    ws['A2'] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    # Row 4: Column headers — Gold Standard 8 columns, 003366 fill
    er_headers = [
        ("Evidence ID", 15),
        ("Assessment Area", 28),
        ("Evidence Type", 22),
        ("Description", 45),
        ("Location/Path", 40),
        ("Date Collected", 16),
        ("Collected By", 20),
        ("Verification Status", 22),
    ]
    for col_idx, (header, width) in enumerate(er_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 5: F2F2F2 sample row — realistic example values
    _er_grey = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _er_yell = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    sample_vals = [
        "EV-001", "Log Source Inventory",
        "Configuration file", "SIEM log source configuration export showing all registered sources",
        "/evidence/a815.1/siem-sources.pdf", "01.01.2026", "Security Analyst", "Verified",
    ]
    for col_idx, val in enumerate(sample_vals, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = _er_grey
        cell.border = border
        cell.font = Font(name='Calibri', size=10, color='808080', italic=True)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.cell(row=5, column=6).number_format = 'DD.MM.YYYY'

    # Rows 6-105: 100 empty FFFFCC data rows (no pre-filled values)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = _er_yell
            cell.border = border
        ws.cell(row=r, column=6).number_format = 'DD.MM.YYYY'

    # Dropdowns — start after sample row (row 6+)
    validations = []

    dv_ev_type = DataValidation(
        type="list",
        formula1='"Log sample,Configuration file,Screenshot,SIEM query result,Documentation,Policy document,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    dv_ev_type.add("C6:C105")
    validations.append(dv_ev_type)

    dv_verification = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    dv_verification.add("H6:H105")
    validations.append(dv_verification)

    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 14: SUMMARY DASHBOARD SHEET (PART 1)
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard — Gold Standard TABLE 1/2/3 format."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    _d9d9d9 = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    _blue = Font(name='Calibri', size=10, color='000000')
    _bold = Font(name='Calibri', size=10, bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    # Row 1: A1:G1 header
    ws.merge_cells('A1:G1')
    ws['A1'] = "{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    # Row 2: A2:G2 italic subtitle — LEFT aligned
    ws.merge_cells('A2:G2')
    ws['A2'] = ("ISO/IEC 27001:2022 A.8.15 requires event logs recording user activities, "
                "exceptions, faults and security events to be produced. This assessment "
                "evaluates log source completeness across all in-scope systems.")
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    # Row 3: blank

    # ── TABLE 1 ──────────────────────────────────────────────────────────────
    row = 4
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "TABLE 1 \u2014 COMPLIANCE OVERVIEW BY ASSESSMENT AREA"
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row}'].border = border
    row += 1

    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = _d9d9d9
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row += 1

    table1_rows = [
        ("System Inventory Completeness",
         "=COUNTIF('System Inventory'!Q6:Q55,\"Compliant\")",
         "=COUNTIF('System Inventory'!Q6:Q55,\"Partial\")",
         "=COUNTIF('System Inventory'!Q6:Q55,\"Non-Compliant\")",
         "=0"),
        ("Authentication Event Logging",
         "=COUNTIF('Authentication Logging'!P6:P55,100)",
         "=COUNTIFS('Authentication Logging'!P6:P55,\">0\",'Authentication Logging'!P6:P55,\"<100\")",
         "=COUNTIFS('Authentication Logging'!A6:A55,\"<>\",'Authentication Logging'!P6:P55,0)",
         "=0"),
        ("Authorisation & Access Logging",
         "=COUNTIF('Authorisation & Access'!P6:P55,100)",
         "=COUNTIFS('Authorisation & Access'!P6:P55,\">0\",'Authorisation & Access'!P6:P55,\"<100\")",
         "=COUNTIFS('Authorisation & Access'!A6:A55,\"<>\",'Authorisation & Access'!P6:P55,0)",
         "=0"),
        ("Administrative Activity Logging",
         "=COUNTIF('Administrative Activity'!P6:P55,100)",
         "=COUNTIFS('Administrative Activity'!P6:P55,\">0\",'Administrative Activity'!P6:P55,\"<100\")",
         "=COUNTIFS('Administrative Activity'!A6:A55,\"<>\",'Administrative Activity'!P6:P55,0)",
         "=0"),
        ("Security Event Logging",
         "=COUNTIF('Security Event Logging'!P6:P55,100)",
         "=COUNTIFS('Security Event Logging'!P6:P55,\">0\",'Security Event Logging'!P6:P55,\"<100\")",
         "=COUNTIFS('Security Event Logging'!A6:A55,\"<>\",'Security Event Logging'!P6:P55,0)",
         "=0"),
        ("Application & Database Logging",
         "=COUNTIF('Application & Database'!P6:P55,100)",
         "=COUNTIFS('Application & Database'!P6:P55,\">0\",'Application & Database'!P6:P55,\"<100\")",
         "=COUNTIFS('Application & Database'!A6:A55,\"<>\",'Application & Database'!P6:P55,0)",
         "=0"),
        ("Network Device Logging",
         "=COUNTIF('Network Device Logging'!P6:P55,100)",
         "=COUNTIFS('Network Device Logging'!P6:P55,\">0\",'Network Device Logging'!P6:P55,\"<100\")",
         "=COUNTIFS('Network Device Logging'!A6:A55,\"<>\",'Network Device Logging'!P6:P55,0)",
         "=0"),
        ("Gap Remediation",
         "=COUNTIF('Gap Analysis'!K6:K55,\"*Resolved*\")",
         "=COUNTIF('Gap Analysis'!K6:K55,\"*In Progress*\")",
         "=COUNTIF('Gap Analysis'!K6:K55,\"*Open*\")",
         "=COUNTIF('Gap Analysis'!K6:K55,\"*Deferred*\")"),
    ]

    data_start = row
    for area_name, comp_f, part_f, nc_f, na_f in table1_rows:
        ws.cell(row=row, column=1, value=area_name).font = _bold
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=f"=C{row}+D{row}+E{row}+F{row}").font = _blue
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3, value=comp_f).font = _blue
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=4, value=part_f).font = _blue
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=5, value=nc_f).font = _blue
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=6, value=na_f).font = _blue
        ws.cell(row=row, column=6).border = border
        pct_cell = ws.cell(row=row, column=7, value=f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))")
        pct_cell.font = _blue
        pct_cell.number_format = '0.0%'
        pct_cell.border = border
        row += 1

    data_end = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=row, column=1).fill = _d9d9d9
    ws.cell(row=row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        total_cell = ws.cell(row=row, column=col_idx,
                             value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        total_cell.font = Font(name='Calibri', size=10, bold=True)
        total_cell.fill = _d9d9d9
        total_cell.border = border
    pct_total = ws.cell(row=row, column=7,
                        value=f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))")
    pct_total.font = Font(name='Calibri', size=10, bold=True)
    pct_total.fill = _d9d9d9
    pct_total.number_format = '0.0%'
    pct_total.border = border
    row += 2

    # ── TABLE 2 ──────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "TABLE 2 \u2014 KEY METRICS"
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row}'].border = border
    row += 1

    t2_headers = ["Metric", "Value", "Target", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=hdr if hdr else None)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = _d9d9d9
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    table2_rows = [
        ("Total systems in scope", "=COUNTA('System Inventory'!B6:B55)", "All ISMS-scope systems"),
        ("Systems with logging enabled", "=COUNTIF('System Inventory'!O6:O55,\"\u2705 Yes\")", "100%"),
        ("Systems forwarding to SIEM", "=COUNTIF('System Inventory'!P6:P55,\"\u2705 Yes\")", "100%"),
        ("P1-Critical systems compliant",
         "=COUNTIFS('System Inventory'!I6:I55,\"P1-Critical\",'System Inventory'!Q6:Q55,\"Compliant\")",
         "100%"),
        ("Total gaps identified", "=COUNTA('Gap Analysis'!B6:B55)", "0"),
        ("Open gaps requiring attention", "=COUNTIF('Gap Analysis'!K6:K55,\"*Open*\")", "0"),
    ]

    for metric_name, metric_formula, target in table2_rows:
        ws.cell(row=row, column=1, value=metric_name).font = _bold
        ws.cell(row=row, column=1).border = border
        val_cell = ws.cell(row=row, column=2, value=metric_formula)
        val_cell.font = _blue
        val_cell.border = border
        if '%' in target:
            val_cell.number_format = '0.0%'
        ws.cell(row=row, column=3, value=target).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=3).border = border
        for col in range(4, 8):
            ws.cell(row=row, column=col).border = border
        row += 1
    row += 1

    # ── TABLE 3 ──────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "TABLE 3 \u2014 CRITICAL FINDINGS REQUIRING ATTENTION"
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row}'].border = border
    row += 1

    t3_headers = ["Finding Type", "Risk Level", "Associated Sheet", "Recommended Action", "ISO Clause"]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = _d9d9d9
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row += 1

    _yell = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    table3_rows = [
        ("System not logging required events", "Critical",
         "System Inventory / Assessment sheets",
         "Enable logging per system type and configure SIEM forwarding immediately",
         "A.8.15(a-j)"),
        ("Authentication events not logged (login failures)", "Critical",
         "Authentication Logging",
         "Configure authentication event logging for all in-scope systems",
         "A.8.15(a)"),
        ("Privileged activity not logged", "Critical",
         "Administrative Activity",
         "Enable administrative action logging (sudo, admin console, service accounts)",
         "A.8.15(d)(e)"),
        ("Logs not forwarded to central SIEM", "High",
         "System Inventory col P",
         "Configure log forwarders; centralisation is an explicit ISO 27002:2022 requirement",
         "A.8.15"),
        ("Open gap past target remediation date", "High",
         "Gap Analysis",
         "Review overdue gaps and escalate to Information Security Manager",
         "A.8.15"),
    ]

    for finding, risk, sheet_ref, action, clause in table3_rows:
        for col_idx, val in enumerate([finding, risk, sheet_ref, action, clause], start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = _yell
            cell.border = border
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    # Freeze rows 1-3
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 15: SUMMARY DASHBOARD SHEET (PART 2 - GAP SUMMARY)
# ============================================================================

def add_gap_summary_to_dashboard(ws, styles):
    """No-op — gap summary now integrated into Summary Dashboard TABLE 1/2/3."""
    pass


# ============================================================================
# SECTION 16: SUMMARY DASHBOARD CHARTS (OPTIONAL)
# ============================================================================

def add_charts_to_dashboard(ws):
    """No-op — charts removed; Summary Dashboard uses TABLE 1/2/3 format."""
    pass

# ============================================================================
# SECTION 17: APPROVAL & SIGN-OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval & Sign-Off sheet (golden standard)."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Header A1:E1
    ws.merge_cells('A1:E1')
    ws['A1'] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border

    # Row 2: Italic subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws['A2'].font = Font(italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ASSESSMENT SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G11"),
        ("Assessment Status:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        if value == "":
            ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        if label == "Assessment Status:":
            status_row = row
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # Status dropdown
    validations = []
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    if status_row:
        dv_status.add(ws[f'B{status_row}'])
    validations.append(dv_status)

    # 3 approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2  # gap before first approver
    for title, colour in approvers:
        # Banner
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=11)
        ws[f'A{row}'].fill = PatternFill(start_color=colour, end_color=colour, fill_type='solid')
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f'A{row}'] = field
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f'A{row}'] = "FINAL DECISION:"
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    dv_decision.add(ws[f'B{row}'])
    validations.append(dv_decision)

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "NEXT REVIEW DETAILS"
    ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = 'A3'

# ============================================================================
# SECTION 18: CONDITIONAL FORMATTING
# ============================================================================

def apply_conditional_formatting(wb):
    """
    Apply conditional formatting rules across all relevant sheets.
    
    "Color is a power which directly influences the soul." - Kandinsky
    Let's use it to influence compliance behaviour!
    """
    
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    
    # Define fills for conditional formatting
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # ==================== SYSTEM INVENTORY SHEET ====================
    ws = wb['System Inventory']

    # Compliance Status column (Q)
    ws.conditional_formatting.add('Q6:Q55',
        CellIsRule(operator='equal', formula=['"Compliant"'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('Q6:Q55',
        CellIsRule(operator='equal', formula=['"Partial"'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('Q6:Q55',
        CellIsRule(operator='equal', formula=['"Non-Compliant"'], stopIfTrue=True, fill=red_fill))

    # ==================== LOG EVENT TYPES SHEET ====================
    ws = wb['Log Event Types by System']

    # Event Types Completeness (R) - percentage-based
    ws.conditional_formatting.add('R6:R55',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.9'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('R6:R55',
        CellIsRule(operator='between', formula=['0.7', '0.89'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('R6:R55',
        CellIsRule(operator='lessThan', formula=['0.7'], stopIfTrue=True, fill=red_fill))

    # ==================== ASSESSMENT SHEETS ====================
    assessment_sheets = [
        'Authentication Logging',
        'Authorisation & Access',
        'Administrative Activity',
        'Security Event Logging',
        'Application & Database',
        'Network Device Logging'
    ]

    for sheet_name in assessment_sheets:
        ws = wb[sheet_name]

        # Compliance Score column (P) - percentage values
        ws.conditional_formatting.add('P6:P55',
            CellIsRule(operator='greaterThanOrEqual', formula=['95'], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add('P6:P55',
            CellIsRule(operator='between', formula=['80', '94'], stopIfTrue=True, fill=yellow_fill))
        ws.conditional_formatting.add('P6:P55',
            CellIsRule(operator='lessThan', formula=['80'], stopIfTrue=True, fill=red_fill))

    # ==================== GAP ANALYSIS SHEET ====================
    ws = wb['Gap Analysis']

    # Priority column (G)
    ws.conditional_formatting.add('G6:G55',
        CellIsRule(operator='equal', formula=['"Critical"'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('G6:G55',
        CellIsRule(operator='equal', formula=['"High"'], stopIfTrue=True,
                   fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
    ws.conditional_formatting.add('G6:G55',
        CellIsRule(operator='equal', formula=['"Medium"'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('G6:G55',
        CellIsRule(operator='equal', formula=['"Low"'], stopIfTrue=True, fill=green_fill))

    # Status column (K)
    ws.conditional_formatting.add('K6:K55',
        CellIsRule(operator='equal', formula=['"Open"'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('K6:K55',
        CellIsRule(operator='equal', formula=['"In Progress"'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('K6:K55',
        CellIsRule(operator='equal', formula=['"Resolved"'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('K6:K55',
        CellIsRule(operator='equal', formula=['"Deferred"'], stopIfTrue=True, fill=gray_fill))

    # Target Date column (J) - highlight overdue dates
    # Note: Date-based conditional formatting needs special handling in openpyxl
    # This is a simplified version
    from openpyxl.formatting.rule import FormulaRule
    ws.conditional_formatting.add('J6:J55',
        FormulaRule(formula=['AND(J6<TODAY(),K6<>"Resolved")'], stopIfTrue=True, fill=red_fill))
    
# ============================================================================
# SECTION 19: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    "The secret to getting ahead is getting started." - Mark Twain
    Let's get this ISMS implementation started properly!
    """
    
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.15.1 - Log Source Inventory Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.15: Logging")
    logger.info("=" * 78)
    logger.info("")
    
    # Create workbook and styles
    logger.info("[1/15] Creating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    # Create each sheet
    logger.info("[2/15] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    
    logger.info("[3/15] Creating System Inventory...")
    create_system_inventory_sheet(wb["System Inventory"], styles)
    
    logger.info("[4/15] Creating Log Event Types by System...")
    create_log_event_types_sheet(wb["Log Event Types by System"], styles)
    
    logger.info("[5/15] Creating Authentication Logging Assessment...")
    create_authentication_logging_sheet(wb["Authentication Logging"], styles)
    
    logger.info("[6/15] Creating Authorisation & Access Logging Assessment...")
    create_authorisation_logging_sheet(wb["Authorisation & Access"], styles)
    
    logger.info("[7/15] Creating Administrative Activity Logging Assessment...")
    create_administrative_activity_sheet(wb["Administrative Activity"], styles)
    
    logger.info("[8/15] Creating Security Event Logging Assessment...")
    create_security_event_logging_sheet(wb["Security Event Logging"], styles)
    
    logger.info("[9/15] Creating Application & Database Logging Assessment...")
    create_application_database_logging_sheet(wb["Application & Database"], styles)
    
    logger.info("[10/15] Creating Network Device Logging Assessment...")
    create_network_device_logging_sheet(wb["Network Device Logging"], styles)
    
    logger.info("[11/15] Creating Gap Analysis...")
    create_gap_analysis_sheet(wb["Gap Analysis"], styles)
    
    logger.info("[12/15] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    
    logger.info("[13/15] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    add_gap_summary_to_dashboard(wb["Summary Dashboard"], styles)
    try:
        add_charts_to_dashboard(wb["Summary Dashboard"])
    except Exception as e:
        logger.info(f"    Note: Charts can be added manually in Excel")
    
    logger.info("[14/15] Creating Approval & Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    
    logger.info("[15/15] Applying conditional formatting...")
    apply_conditional_formatting(wb)
    
    # Generate filename with today's date
    filename = f"ISMS-IMP-A.8.15.1_Log_Source_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"

    logger.info("")
    logger.info("Saving workbook...")
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info("")
    logger.info("=" * 78)
    logger.info("\u2705 SUCCESS: Workbook generated successfully!")
    logger.info("=" * 78)
    logger.info("")
    logger.info(f"Filename: {filename}")
    logger.info(f"File size: ~800 KB - 1.5 MB (estimated)")
    logger.info("")
    logger.info("Workbook Structure:")
    logger.info("  ✓ Sheet 1:  Instructions & Legend")
    logger.info("  ✓ Sheet 2:  System Inventory (92 data rows)")
    logger.info("  ✓ Sheet 3:  Log Event Types by System (92 data rows)")
    logger.info("  ✓ Sheet 4:  Authentication Logging Assessment (92 rows)")
    logger.info("  ✓ Sheet 5:  Authorisation & Access Logging (92 rows)")
    logger.info("  ✓ Sheet 6:  Administrative Activity Logging (92 rows)")
    logger.info("  ✓ Sheet 7:  Security Event Logging (92 rows)")
    logger.info("  ✓ Sheet 8:  Application & Database Logging (92 rows)")
    logger.info("  ✓ Sheet 9:  Network Device Logging (92 rows)")
    logger.info("  ✓ Sheet 10: Gap Analysis (92 gap rows)")
    logger.info("  ✓ Sheet 11: Evidence Register (100 evidence rows)")
    logger.info("  ✓ Sheet 12: Summary Dashboard (with KPIs and gap summary)")
    logger.info("  ✓ Sheet 13: Approval & Sign-Off (multi-level workflow)")
    logger.info("")
    logger.info("Features:")
    logger.info("  ✓ Auto-generated IDs (System, Gap, Evidence)")
    logger.info("  ✓ VLOOKUP formulas for system names")
    logger.info("  ✓ Compliance score calculations")
    logger.info("  ✓ Data validation dropdowns")
    logger.info("  ✓ Conditional formatting (Green/Yellow/Red status)")
    logger.info("  ✓ Date format: DD.MM.YYYY")
    logger.info("  ✓ Frozen panes for easy navigation")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open the workbook in Excel/LibreOffice")
    logger.info("  2. Fill in yellow-highlighted cells (user input required)")
    logger.info("  3. Complete System Inventory sheet first")
    logger.info("  4. Then complete assessment sheets for each system")
    logger.info("  5. Review Summary Dashboard for compliance status")
    logger.info("  6. Document gaps in Gap Analysis sheet")
    logger.info("  7. Collect evidence artifacts in Evidence Register")
    logger.info("  8. Obtain approvals in Approval & Sign-Off sheet")
    logger.info("")
    logger.info("Estimated Completion Time: 4-8 hours (for 50-100 systems)")
    logger.info("")
    logger.info("═" * 78)
    logger.info("Don't cargo cult this assessment - actually read your logs!")
    logger.info("═" * 78)
    logger.info("")


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
