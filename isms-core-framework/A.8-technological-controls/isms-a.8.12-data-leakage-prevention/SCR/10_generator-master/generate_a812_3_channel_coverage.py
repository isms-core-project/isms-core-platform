#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.3 - Channel Coverage Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 3 of 4: Channel Coverage

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data channels, exfiltration vectors, and
protection requirements.

Key customization areas:
1. Email platform configuration (M365, Google Workspace, on-prem)
2. Cloud storage services (specific to your approved services)
3. Endpoint control capabilities (USB, print per your DLP tools)
4. Network monitoring scope (aligned with your architecture)
5. Mobile device management (MDM/MAM specific to your deployment)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates comprehensive channel coverage assessment workbook for systematic
evaluation of DLP protection across all data exfiltration channels against
ISO 27001:2022 Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• Email channel DLP protection (inbound/outbound, internal)
• Web/cloud channel coverage (browser uploads, cloud storage sync)
• Endpoint channel controls (USB, clipboard, print, screen capture)
• Network channel monitoring (file transfers, protocols)
• Application channel security (SaaS, APIs, database exports)
• Mobile channel protection (MDM/MAM, mobile email)
• Channel coverage metrics and gap analysis
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.3_Channel_Coverage_YYYYMMDD.xlsx

Sheets (12 total):
1. Instructions_Legend - Assessment guidance and metadata
2. Channel_Overview - Summary of all 6 channels and protection tiers
3. Email_Channel - Email DLP deployment and coverage assessment
4. Web_Cloud_Channel - Web/cloud DLP deployment and coverage
5. Endpoint_Channel - Endpoint DLP (USB, print, clipboard, etc.)
6. Network_Channel - Network DLP deployment and coverage
7. Application_Channel - Application-level DLP controls
8. Mobile_Channel - Mobile DLP (MDM/MAM) deployment
9. Coverage_Metrics - Channel coverage percentage calculations
10. Gap_Analysis - Channel protection gaps (40 rows)
11. Evidence_Register - Audit evidence tracking (100 rows)
12. Summary_Dashboard - Compliance metrics and KPIs

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_3_channel_coverage.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.3_Channel_Coverage_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organization-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Assess DLP coverage for each data exfiltration channel
    5. Calculate channel coverage percentages
    6. Document DLP policies and actions per channel
    7. Collect and document evidence (Evidence_Register sheet)
    8. Complete gap analysis for unprotected channels
    9. Obtain management approval (Summary_Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 3 of 4 in comprehensive DLP evaluation framework
    Focus: Channel protection coverage across all data exfiltration vectors
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.2 (Channel Protection Requirements)
    Policy:         ISMS-POL-A.8.12-S5.A (DLP Channel Standards)
    Implementation: ISMS-IMP-A.8.12.3 (Channel Coverage Implementation Guide)
    Dashboard:      ISMS-IMP-A.8.12.5 (Compliance Dashboard)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py
       python3 generate_a812_2_data_classification.py
       python3 generate_a812_3_channel_coverage.py         ← YOU ARE HERE
       python3 generate_a812_4_monitoring_response.py
    
    2. Complete assessments (manual - security team, network team)
    
    3. Normalize filenames for dashboard linking:
       python3 normalize_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Channel Assessment → Normalize → Dashboard → Audit Evidence

Critical Prerequisites:
    • Domain 1 (Infrastructure) completed - need DLP technology inventory
    • Domain 2 (Classification) completed - need data classification schema
    • Channel DLP policies require both infrastructure AND data classification

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    3 of 4 (Channel Coverage)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organization] ISMS Implementation Team
License:              [Organization License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - Swiss FADP (Federal Act on Data Protection)
    - EU GDPR (General Data Protection Regulation)
    - NIST SP 800-53 (Security and Privacy Controls - SC-7, SC-8)
    - CIS Controls v8.1 (Center for Internet Security - Controls 3, 13)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Channel Protection Tiers (from ISMS-POL-A.8.12-S2.2):
    Tier 1 (Critical - Months 1-3):
        • Email (inbound/outbound/internal)
        • Web/Browser uploads
        • USB/Removable media
    
    Tier 2 (High - Months 4-6):
        • Cloud storage sync (Dropbox, OneDrive, Google Drive)
        • Mobile devices (BYOD, corporate)
    
    Tier 3 (Medium - Months 7-9):
        • Network file shares (SMB, NFS)
        • Printers (local, network)
    
    Tier 4 (Low - Months 10-12):
        • Bluetooth
        • Optical drives (CD/DVD)
        • Legacy protocols (FTP, Telnet)

Assessment Scope - Six Data Exfiltration Channels:
    1. Email Channel:
       • Outbound email attachments and body content
       • Inbound email (data import risks)
       • Internal email (insider threat)
       • Webmail (Gmail, Outlook.com)
    
    2. Web/Cloud Channel:
       • Browser file uploads (HTTP POST, WebDAV)
       • Cloud storage sync clients
       • Web-based file sharing (WeTransfer, SendAnywhere)
       • SaaS application data exports
    
    3. Endpoint Channel:
       • USB/removable media (flash drives, external HDD)
       • Clipboard operations (copy/paste to external apps)
       • Print spooler (local and network printers)
       • Screen capture tools
       • Bluetooth file transfers
    
    4. Network Channel:
       • File transfer protocols (FTP, SFTP, SCP)
       • Network shares (SMB, NFS, WebDAV)
       • Database exports (SQL dumps, CSV exports)
       • Backup systems (unauthorized backups)
    
    5. Application Channel:
       • SaaS application APIs
       • Database connection exports
       • ERP/CRM data exports
       • Source code repositories (Git push)
    
    6. Mobile Channel:
       • Mobile email clients
       • Mobile cloud sync apps
       • SMS/MMS (data via messaging)
       • Mobile app data sharing

Coverage Calculation Methodology:
    Channel Coverage % = (Protected Users/Endpoints / Total Users/Endpoints) × 100
    
    Overall DLP Coverage = Weighted average across channels:
    • Tier 1 channels: 40% weight
    • Tier 2 channels: 30% weight
    • Tier 3 channels: 20% weight
    • Tier 4 channels: 10% weight

Data Classification:
    Generated workbooks contain sensitive organizational security information.
    Handle according to [Organization]'s data classification policy.
    Recommended classification: [Organization] Internal/Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: channel coverage %, protection tiers, gap remediation.

Review Cycle:
    Quarterly: Update channel coverage metrics
    Annually: Complete full channel assessment
    Ad-hoc: When new channels identified or infrastructure changes

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.3"
WORKBOOK_NAME = "Channel Coverage Assessment"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: CONSTANTS & CONFIGURATION
# ============================================================================

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.3"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.2"
ASSESSMENT_AREA = "Channel Coverage Assessment"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "E7E6E6"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "B4C7E7"         # Light blue (\u1F4CB Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 40


# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    sheets = [
        "Instructions_Legend",
        "Channel_Overview",
        "Email_Channel",
        "Web_Cloud_Channel",
        "Endpoint_Channel",
        "Network_Channel",
        "Application_Channel",
        "Mobile_Channel",
        "Coverage_Metrics",
        "Gap_Analysis",
        "Evidence_Register",
        "Summary_Dashboard",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    styles = {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "font": Font(name="Calibri", size=10, italic=True),
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "normal_cell": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "bold_label": {
            "font": Font(name="Calibri", size=10, bold=True),
            "alignment": Alignment(horizontal="left", vertical="center")
        }
    }
    
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell by creating NEW style instances.
    NEVER reuse Font/Fill/Border objects across cells.
    """
    if "font" in style_dict:
        cell.font = Font(**{k: v for k, v in style_dict["font"].__dict__.items() if not k.startswith('_')})
    if "fill" in style_dict:
        cell.fill = PatternFill(**{k: v for k, v in style_dict["fill"].__dict__.items() if not k.startswith('_')})
    if "alignment" in style_dict:
        cell.alignment = Alignment(**{k: v for k, v in style_dict["alignment"].__dict__.items() if not k.startswith('_')})
    if "border" in style_dict:
        cell.border = Border(**{k: v for k, v in style_dict["border"].__dict__.items() if not k.startswith('_')})


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """
    Create data validation objects.
    MUST be added to worksheet.add_data_validation() and then cells added to validation.
    """
    return {
        "yes_no_partial": DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            error="Invalid value. Select from dropdown.",
            errorTitle="Invalid Entry"
        ),
        "channel_type": DataValidation(
            type="list",
            formula1='"Email,Web,Endpoint,Network,App,Mobile"',
            allow_blank=False,
            showDropDown=True
        ),
        "risk_level": DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "exfiltration_risk": DataValidation(
            type="list",
            formula1='"Very High,High,Medium,Low,Very Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "status": DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Complete,Blocked"',
            allow_blank=False,
            showDropDown=True
        ),
        "approval_status": DataValidation(
            type="list",
            formula1='"Approved,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
        "email_system_type": DataValidation(
            type="list",
            formula1='"SMTP,M365,Gmail,Webmail,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_mode": DataValidation(
            type="list",
            formula1='"Inline,Monitor,Cloud-Native,CASB"',
            allow_blank=False,
            showDropDown=True
        ),
        "policy_action": DataValidation(
            type="list",
            formula1='"Allow,Alert,Block,Quarantine,Encrypt"',
            allow_blank=False,
            showDropDown=True
        ),
        "encrypted_handling": DataValidation(
            type="list",
            formula1='"Decrypt,Quarantine,Allow,Block"',
            allow_blank=False,
            showDropDown=True
        ),
        "service_type": DataValidation(
            type="list",
            formula1='"Cloud Storage,SaaS,Web Upload,Code Repo,Social Media"',
            allow_blank=False,
            showDropDown=True
        ),
        "protection_method": DataValidation(
            type="list",
            formula1='"Proxy,CASB,Cloud-Native,Endpoint"',
            allow_blank=False,
            showDropDown=True
        ),
        "endpoint_type": DataValidation(
            type="list",
            formula1='"Windows Desktop,macOS,Linux,VDI,Thin Client"',
            allow_blank=False,
            showDropDown=True
        ),
        "usb_control": DataValidation(
            type="list",
            formula1='"Block All,Allow Approved,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "print_control": DataValidation(
            type="list",
            formula1='"Block,Watermark,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "clipboard_control": DataValidation(
            type="list",
            formula1='"Block,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "app_control": DataValidation(
            type="list",
            formula1='"Whitelist,Blacklist,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "network_protocol": DataValidation(
            type="list",
            formula1='"SMB/CIFS,FTP,SFTP,NFS,SCP,WebDAV,Rsync"',
            allow_blank=False,
            showDropDown=True
        ),
        "detection_method": DataValidation(
            type="list",
            formula1='"Inline,TAP,SPAN,Endpoint,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "application_type": DataValidation(
            type="list",
            formula1='"Database,API,Reporting,CRM,ERP,BI,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "export_capability": DataValidation(
            type="list",
            formula1='"Yes,No,Limited"',
            allow_blank=False,
            showDropDown=True
        ),
        "dlp_control_method": DataValidation(
            type="list",
            formula1='"DAM,API Gateway,App Control,Endpoint,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "export_action": DataValidation(
            type="list",
            formula1='"Allow,Alert,Block,Approval Required"',
            allow_blank=False,
            showDropDown=True
        ),
        "device_ownership": DataValidation(
            type="list",
            formula1='"Corporate,BYOD"',
            allow_blank=False,
            showDropDown=True
        ),
        "mobile_platform": DataValidation(
            type="list",
            formula1='"iOS,Android,Windows Mobile"',
            allow_blank=False,
            showDropDown=True
        ),
        "evidence_type": DataValidation(
            type="list",
            formula1='"Config,Screenshot,Log,Report,Policy,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "verification_status": DataValidation(
            type="list",
            formula1='"Verified,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        )
    }


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 1 of 4)
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ISMS Control {WORKBOOK_ID} - {ASSESSMENT_AREA}"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    # Subheader
    ws.merge_cells('A2:H2')
    ws['A2'] = f"ISO/IEC 27001:2022 Control {CONTROL_ID} - Data Leakage Prevention"
    apply_style(ws['A2'], styles["subheader"])
    
    # Document Info
    info = [
        ("Document ID:", WORKBOOK_ID),
        ("Assessment Area:", ASSESSMENT_AREA),
        ("Related Policy:", RELATED_POLICY),
        ("Version:", WORKBOOK_VERSION),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row = 4
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "USER INPUT" in value:
            apply_style(ws[f'B{row}'], styles["input_cell"])
        row += 1
    
    # HOW TO USE THIS WORKBOOK section
    row += 2
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    instructions = [
        "1. Complete Channel_Overview first for high-level assessment of all 6 channels",
        "2. Complete each channel-specific sheet (Email, Web, Endpoint, Network, Application, Mobile)",
        "3. Fill ALL yellow-highlighted cells with your organization's information",
        "4. Use dropdown menus where provided (do not type free-form text in dropdown cells)",
        "5. Document DLP deployment status, coverage %, and policy actions per channel",
        "6. Review Coverage_Metrics sheet for channel coverage calculations",
        "7. Identify gaps in Gap_Analysis sheet and create remediation plans",
        "8. Provide evidence IDs for all assessments (Evidence ID format: A812-3-[CHANNEL]-[###])",
        "9. Review Summary Dashboard for overall compliance score",
        "10. Obtain CISO/DPO approval before finalizing",
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        row += 1
    
    # LEGEND section
    row += 2
    ws[f'A{row}'] = "LEGEND - RESPONSE VALUES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    legend = [
        ("Yes", "Fully implemented and documented", COLOR_COMPLETE),
        ("No", "Not implemented", COLOR_MISSING),
        ("Partial", "Partially implemented (explain in notes)", COLOR_PARTIAL),
        ("Planned", "Scheduled for implementation (provide target date)", COLOR_PLANNED),
        ("N/A", "Not applicable (provide justification)", "FFFFFF"),
    ]
    
    for value, description, color in legend:
        ws[f'A{row}'] = value
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'B{row}'] = description
        row += 1
    
    # CHANNEL PROTECTION TIERS section
    row += 2
    ws[f'A{row}'] = "CHANNEL PROTECTION TIERS (from ISMS-POL-A.8.12-S2.2)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    tiers = [
        ("Tier 1 (Critical)", "Email, Web, USB", "Months 1-3", "Highest exfiltration risk"),
        ("Tier 2 (High)", "Cloud, Mobile", "Months 4-6", "Growing attack surface"),
        ("Tier 3 (Medium)", "Network shares, Print", "Months 7-9", "Lower volume"),
        ("Tier 4 (Low)", "Bluetooth, Optical drives", "Months 10-12", "Niche vectors"),
    ]
    
    ws[f'A{row}'] = "Tier"
    ws[f'B{row}'] = "Channels"
    ws[f'C{row}'] = "Timeline"
    ws[f'D{row}'] = "Rationale"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
    row += 1
    
    for tier_name, channels, timeline, rationale in tiers:
        ws[f'A{row}'] = tier_name
        ws[f'B{row}'] = channels
        ws[f'C{row}'] = timeline
        ws[f'D{row}'] = rationale
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_DESCRIPTION
    
    # Freeze panes at row 3
    ws.freeze_panes = 'A3'


def create_channel_overview(ws, styles, validations):
    """Create Channel Overview summary sheet."""
    
    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "CHANNEL COVERAGE OVERVIEW"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Summary of DLP protection across all data egress channels"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment info
    ws['A3'] = "Assessment Date:"
    ws['A3'].font = Font(bold=True)
    apply_style(ws['B3'], styles["input_cell"])
    ws['D3'] = "Completed By:"
    ws['D3'].font = Font(bold=True)
    apply_style(ws['E3'], styles["input_cell"])
    
    # Column headers
    headers = [
        "Channel Name", "Priority Tier", "DLP Deployed?", "Coverage %", 
        "Users Protected", "Total Users", "Policy Action", "Gap Status", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated channel data
    channels = [
        ("Email", "Tier 1", "[Input]", "=Email_Channel!B25", "[Input]", "[Input]", "[Input]", "=IF(D6>=90,\"\u2705 Compliant\",IF(D6>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Web/Cloud", "Tier 1", "[Input]", "=Web_Cloud_Channel!B25", "[Input]", "[Input]", "[Input]", "=IF(D7>=90,\"\u2705 Compliant\",IF(D7>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Endpoint", "Tier 1", "[Input]", "=Endpoint_Channel!L30", "[Input]", "[Input]", "[Input]", "=IF(D8>=90,\"\u2705 Compliant\",IF(D8>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Network", "Tier 3", "[Input]", "=Network_Channel!E20", "[Input]", "[Input]", "[Input]", "=IF(D9>=75,\"\u2705 Compliant\",IF(D9>=60,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Application", "Tier 2", "[Input]", "=Application_Channel!E20", "[Input]", "[Input]", "[Input]", "=IF(D10>=80,\"\u2705 Compliant\",IF(D10>=65,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Mobile", "Tier 2", "[Input]", "=Mobile_Channel!L20", "[Input]", "[Input]", "[Input]", "=IF(D11>=90,\"\u2705 Compliant\",IF(D11>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
    ]
    
    for row_idx, channel_data in enumerate(channels, start=6):
        for col_idx, value in enumerate(channel_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if "[Input]" in str(value) and col_idx not in [4, 8]:  # Not formula columns
                apply_style(cell, styles["input_cell"])
            else:
                apply_style(cell, styles["normal_cell"])
    
    # Add data validations
    ws.add_data_validation(validations["yes_no_partial"])
    for row in range(6, 12):
        validations["yes_no_partial"].add(ws[f'C{row}'])
    
    ws.add_data_validation(validations["policy_action"])
    for row in range(6, 12):
        validations["policy_action"].add(ws[f'G{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


# END OF PART 1

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 2 of 4)
# ============================================================================

def create_email_channel(ws, styles, validations):
    """Create Email Channel assessment sheet."""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "EMAIL CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:L2')
    ws['A2'] = "Evaluate DLP coverage for SMTP, M365, Gmail, Webmail"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:L3')
    ws['A3'] = "Does your organization have DLP protection for all email systems?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Email System", "System Type", "DLP Solution", "Deployment Mode",
        "Content Inspection", "Attachment Scanning", "Encrypted Email Handling",
        "Policy Action", "External Email Protected", "Users Covered",
        "SIEM Integration", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (gray rows)
    examples = [
        ("Exchange Online", "M365", "Microsoft Purview", "Cloud-Native", "Yes", "Yes", "Decrypt", "Block", "Yes", "850", "Yes", "A812-3-EMAIL-001"),
        ("Gmail", "Gmail", "Google Workspace DLP", "Cloud-Native", "Yes", "Yes", "Quarantine", "Alert", "Yes", "120", "Yes", "A812-3-EMAIL-002"),
        ("SMTP Gateway", "SMTP", "Forcepoint DLP", "Inline", "Yes", "Yes", "Decrypt", "Block", "Yes", "900", "Yes", "A812-3-EMAIL-003"),
    ]
    
    for row_idx, example_data in enumerate(examples, start=6):
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
    
    # Blank input rows
    for row in range(9, 25):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
    
    # Summary row
    ws['A25'] = "Email Channel Coverage %"
    ws['A25'].font = Font(bold=True)
    ws['B25'] = "=(COUNTIF(E6:E24,\"Yes\")/COUNTA(A6:A24))*100"
    ws['B25'].number_format = '0.0"%"'
    
    # Add data validations
    ws.add_data_validation(validations["email_system_type"])
    for row in range(6, 25):
        validations["email_system_type"].add(ws[f'B{row}'])
    
    ws.add_data_validation(validations["deployment_mode"])
    for row in range(6, 25):
        validations["deployment_mode"].add(ws[f'D{row}'])
    
    ws.add_data_validation(validations["yes_no_partial"])
    for row in range(6, 25):
        for col in ['E', 'F', 'I', 'K']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])
    
    ws.add_data_validation(validations["encrypted_handling"])
    for row in range(6, 25):
        validations["encrypted_handling"].add(ws[f'G{row}'])
    
    ws.add_data_validation(validations["policy_action"])
    for row in range(6, 25):
        validations["policy_action"].add(ws[f'H{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_web_cloud_channel(ws, styles, validations):
    """Create Web/Cloud Channel assessment sheet."""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "WEB/CLOUD CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Evaluate DLP coverage for web uploads, cloud storage, SaaS applications"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:K3')
    ws['A3'] = "Does your organization protect against data exfiltration via web and cloud services?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Service/Application", "Service Type", "DLP Solution", "Protection Method",
        "SSL/TLS Inspection", "Upload Blocking", "Download Monitoring",
        "Policy Action", "Users Covered", "SIEM Integration", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (gray rows)
    examples = [
        ("OneDrive", "Cloud Storage", "Microsoft Purview", "Cloud-Native", "Yes", "Yes", "Yes", "Block", "850", "Yes", "A812-3-WEB-001"),
        ("Dropbox", "Cloud Storage", "Netskope CASB", "CASB", "Yes", "Yes", "Yes", "Alert", "900", "Yes", "A812-3-WEB-002"),
        ("GitHub", "Code Repo", "Forcepoint Web DLP", "Proxy", "Yes", "Yes", "Partial", "Block", "200", "Yes", "A812-3-WEB-003"),
        ("Personal Webmail", "Web Upload", "Zscaler DLP", "Proxy", "Yes", "Yes", "No", "Block", "900", "Yes", "A812-3-WEB-004"),
    ]
    
    for row_idx, example_data in enumerate(examples, start=6):
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
    
    # Blank input rows
    for row in range(10, 25):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
    
    # Summary row
    ws['A25'] = "Web/Cloud Channel Coverage %"
    ws['A25'].font = Font(bold=True)
    ws['B25'] = "=(COUNTIF(E6:E24,\"Yes\")/COUNTA(A6:A24))*100"
    ws['B25'].number_format = '0.0"%"'
    
    # Add data validations
    ws.add_data_validation(validations["service_type"])
    for row in range(6, 25):
        validations["service_type"].add(ws[f'B{row}'])
    
    ws.add_data_validation(validations["protection_method"])
    for row in range(6, 25):
        validations["protection_method"].add(ws[f'D{row}'])
    
    ws.add_data_validation(validations["yes_no_partial"])
    for row in range(6, 25):
        for col in ['E', 'F', 'G', 'J']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])
    
    ws.add_data_validation(validations["policy_action"])
    for row in range(6, 25):
        validations["policy_action"].add(ws[f'H{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_WIDE
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_endpoint_channel(ws, styles, validations):
    """Create Endpoint Channel assessment sheet."""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "ENDPOINT CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:M2')
    ws['A2'] = "Evaluate endpoint DLP agent coverage and controls"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:M3')
    ws['A3'] = "Does your organization have endpoint DLP agents protecting USB, print, and other local channels?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Endpoint Type", "OS Version", "DLP Agent Installed", "USB Control",
        "Print Control", "Clipboard Control", "Screen Capture Control",
        "Bluetooth Control", "Application Control", "Devices Covered",
        "Total Devices", "Coverage %", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (gray rows)
    examples = [
        ("Windows Desktop", "Win 11", "Yes", "Allow Approved", "Watermark", "Block", "Block", "Monitor", "Whitelist", "850", "900", "=J6/K6*100", "A812-3-EPT-001"),
        ("macOS", "Sonoma 14", "Yes", "Block All", "Watermark", "Monitor", "Monitor", "Monitor", "Whitelist", "120", "150", "=J7/K7*100", "A812-3-EPT-002"),
        ("Linux", "Ubuntu 22.04", "No", "None", "None", "None", "None", "None", "None", "0", "50", "=J8/K8*100", "A812-3-EPT-003"),
        ("VDI", "Win 11", "Yes", "Block All", "Block", "Block", "Block", "Block", "Whitelist", "200", "200", "=J9/K9*100", "A812-3-EPT-004"),
    ]
    
    for row_idx, example_data in enumerate(examples, start=6):
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
    
    # Blank input rows
    for row in range(10, 30):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
            # Add formula for Coverage % column
            if col == 12:
                cell.value = f"=J{row}/K{row}*100"
                cell.number_format = '0.0"%"'
    
    # Summary row
    ws['A30'] = "Endpoint Channel Coverage %"
    ws['A30'].font = Font(bold=True)
    ws['L30'] = "=AVERAGE(L6:L29)"
    ws['L30'].number_format = '0.0"%"'
    
    # Add data validations
    ws.add_data_validation(validations["endpoint_type"])
    for row in range(6, 30):
        validations["endpoint_type"].add(ws[f'A{row}'])
    
    ws.add_data_validation(validations["yes_no_partial"])
    for row in range(6, 30):
        validations["yes_no_partial"].add(ws[f'C{row}'])
    
    ws.add_data_validation(validations["usb_control"])
    for row in range(6, 30):
        validations["usb_control"].add(ws[f'D{row}'])
    
    ws.add_data_validation(validations["print_control"])
    for row in range(6, 30):
        for col in ['E', 'G']:
            validations["print_control"].add(ws[f'{col}{row}'])
    
    ws.add_data_validation(validations["clipboard_control"])
    for row in range(6, 30):
        for col in ['F', 'H']:
            validations["clipboard_control"].add(ws[f'{col}{row}'])
    
    ws.add_data_validation(validations["app_control"])
    for row in range(6, 30):
        validations["app_control"].add(ws[f'I{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_network_channel(ws, styles, validations):
    """Create Network Channel assessment sheet."""
    
    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "NETWORK CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Evaluate network DLP coverage for SMB, FTP, NFS, SCP"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:I3')
    ws['A3'] = "Does your organization monitor/block file transfers via network protocols?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Protocol", "Business Use Case", "DLP Solution", "Detection Method",
        "Content Inspection", "Policy Action", "Encryption Handling",
        "SIEM Integration", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (gray rows)
    examples = [
        ("SMB/CIFS", "Internal file shares", "Forcepoint DLP", "Inline", "Yes", "Alert", "Allow", "Yes", "A812-3-NET-001"),
        ("FTP", "External file transfer", "Network DLP", "TAP", "Yes", "Block", "Decrypt", "Yes", "A812-3-NET-002"),
        ("SFTP", "Secure file transfer", "Endpoint DLP", "Endpoint", "Partial", "Alert", "Allow", "Partial", "A812-3-NET-003"),
    ]
    
    for row_idx, example_data in enumerate(examples, start=6):
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
    
    # Blank input rows
    for row in range(9, 20):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
    
    # Summary row
    ws['A20'] = "Network Channel Coverage %"
    ws['A20'].font = Font(bold=True)
    ws['E20'] = "=(COUNTIF(E6:E19,\"Yes\")/COUNTA(A6:A19))*100"
    ws['E20'].number_format = '0.0"%"'
    
    # Add data validations
    ws.add_data_validation(validations["network_protocol"])
    for row in range(6, 20):
        validations["network_protocol"].add(ws[f'A{row}'])
    
    ws.add_data_validation(validations["detection_method"])
    for row in range(6, 20):
        validations["detection_method"].add(ws[f'D{row}'])
    
    ws.add_data_validation(validations["yes_no_partial"])
    for row in range(6, 20):
        for col in ['E', 'H']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])
    
    policy_action_net = DataValidation(
        type="list",
        formula1='"Allow,Alert,Block,None"',
        allow_blank=False,
        showDropDown=True
    )
    ws.add_data_validation(policy_action_net)
    for row in range(6, 20):
        policy_action_net.add(ws[f'F{row}'])
    
    encrypt_handling_net = DataValidation(
        type="list",
        formula1='"Decrypt,Allow,Block"',
        allow_blank=False,
        showDropDown=True
    )
    ws.add_data_validation(encrypt_handling_net)
    for row in range(6, 20):
        encrypt_handling_net.add(ws[f'G{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


# END OF PART 2

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 3 of 4)
# ============================================================================

def create_application_channel(ws, styles, validations):
    """Create Application Channel assessment sheet."""
    
    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "APPLICATION CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Evaluate DLP controls in databases, APIs, reporting tools, CRM, ERP"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:I3')
    ws['A3'] = "Does your organization have DLP controls for application-level data exports?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Application Name", "Application Type", "Data Export Capability",
        "DLP Control Method", "Bulk Export Detection", "Export Policy Action",
        "Audit Logging", "SIEM Integration", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (gray rows)
    examples = [
        ("Oracle DB", "Database", "Yes", "DAM (Imperva)", "Yes", "Alert", "Yes", "Yes", "A812-3-APP-001"),
        ("Salesforce", "CRM", "Yes", "CASB", "Yes", "Block", "Yes", "Yes", "A812-3-APP-002"),
        ("Power BI", "BI", "Yes", "App Control", "Partial", "Approval Required", "Partial", "Partial", "A812-3-APP-003"),
    ]
    
    for row_idx, example_data in enumerate(examples, start=6):
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
    
    # Blank input rows
    for row in range(9, 20):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
    
    # Summary row
    ws['A20'] = "Application Channel Coverage %"
    ws['A20'].font = Font(bold=True)
    ws['E20'] = "=(COUNTIF(E6:E19,\"Yes\")/COUNTA(A6:A19))*100"
    ws['E20'].number_format = '0.0"%"'
    
    # Add data validations
    ws.add_data_validation(validations["application_type"])
    for row in range(6, 20):
        validations["application_type"].add(ws[f'B{row}'])
    
    ws.add_data_validation(validations["export_capability"])
    for row in range(6, 20):
        validations["export_capability"].add(ws[f'C{row}'])
    
    ws.add_data_validation(validations["dlp_control_method"])
    for row in range(6, 20):
        validations["dlp_control_method"].add(ws[f'D{row}'])
    
    ws.add_data_validation(validations["yes_no_partial"])
    for row in range(6, 20):
        for col in ['E', 'G', 'H']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])
    
    ws.add_data_validation(validations["export_action"])
    for row in range(6, 20):
        validations["export_action"].add(ws[f'F{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = WIDTH_WIDE
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_mobile_channel(ws, styles, validations):
    """Create Mobile Channel assessment sheet."""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "MOBILE CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:M2')
    ws['A2'] = "Evaluate MDM/MAM DLP controls for corporate and BYOD devices"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:M3')
    ws['A3'] = "Does your organization have DLP controls for mobile devices with corporate data access?"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Device Ownership", "Mobile Platform", "MDM/MAM Solution", "Work Profile Enabled",
        "Copy-Paste Control", "Screenshot Control", "Camera Access Control",
        "Personal Cloud Block", "AirDrop/NFC Control", "Devices Enrolled",
        "Total Mobile Devices", "Coverage %", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (gray rows)
    examples = [
        ("Corporate", "iOS", "Microsoft Intune", "Yes", "Block", "Block", "Block", "Yes", "Block", "200", "200", "=J6/K6*100", "A812-3-MOB-001"),
        ("BYOD", "Android", "Workspace ONE", "Yes", "Block", "Monitor", "Monitor", "Partial", "Monitor", "150", "300", "=J7/K7*100", "A812-3-MOB-002"),
        ("Corporate", "iOS", "Jamf Pro", "Yes", "Block", "Watermark", "Block", "Yes", "Block", "50", "50", "=J8/K8*100", "A812-3-MOB-003"),
    ]
    
    for row_idx, example_data in enumerate(examples, start=6):
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
    
    # Blank input rows
    for row in range(9, 20):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
            # Add formula for Coverage % column
            if col == 12:
                cell.value = f"=J{row}/K{row}*100"
                cell.number_format = '0.0"%"'
    
    # Summary row
    ws['A20'] = "Mobile Channel Coverage %"
    ws['A20'].font = Font(bold=True)
    ws['L20'] = "=AVERAGE(L6:L19)"
    ws['L20'].number_format = '0.0"%"'
    
    # Add data validations
    ws.add_data_validation(validations["device_ownership"])
    for row in range(6, 20):
        validations["device_ownership"].add(ws[f'A{row}'])
    
    ws.add_data_validation(validations["mobile_platform"])
    for row in range(6, 20):
        validations["mobile_platform"].add(ws[f'B{row}'])
    
    ws.add_data_validation(validations["yes_no_partial"])
    for row in range(6, 20):
        for col in ['D', 'H']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])
    
    ws.add_data_validation(validations["clipboard_control"])
    for row in range(6, 20):
        for col in ['E', 'G', 'I']:
            validations["clipboard_control"].add(ws[f'{col}{row}'])
    
    ws.add_data_validation(validations["print_control"])
    for row in range(6, 20):
        validations["print_control"].add(ws[f'F{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_coverage_metrics(ws, styles):
    """Create Coverage Metrics sheet."""
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "CHANNEL COVERAGE METRICS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Weighted coverage analysis across all channels"
    apply_style(ws['A2'], styles["subheader"])
    
    # Section 1: Channel Coverage Summary
    ws['A4'] = "SECTION 1: CHANNEL COVERAGE SUMMARY"
    ws['A4'].font = Font(bold=True, size=12)
    
    headers1 = ["Channel", "Target %", "Actual %", "Gap", "Status"]
    for col_idx, header in enumerate(headers1, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    channels_data = [
        ("Email", "95%", "=Email_Channel!B25", "=B6-C6", "=IF(C6>=B6,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Web/Cloud", "90%", "=Web_Cloud_Channel!B25", "=B7-C7", "=IF(C7>=B7,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Endpoint", "95%", "=Endpoint_Channel!L30", "=B8-C8", "=IF(C8>=B8,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Network", "75%", "=Network_Channel!E20", "=B9-C9", "=IF(C9>=B9,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Application", "80%", "=Application_Channel!E20", "=B10-C10", "=IF(C10>=B10,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Mobile", "90%", "=Mobile_Channel!L20", "=B11-C11", "=IF(C11>=B11,\"\u2705 Compliant\",\"\u274C Gap\")"),
    ]
    
    for row_idx, channel_data in enumerate(channels_data, start=6):
        for col_idx, value in enumerate(channel_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["normal_cell"])
            if col_idx in [2, 3, 4]:
                cell.number_format = '0.0"%"'
    
    # Section 2: Weighted Average Coverage
    ws['A13'] = "SECTION 2: WEIGHTED AVERAGE COVERAGE"
    ws['A13'].font = Font(bold=True, size=12)
    
    headers2 = ["Metric", "Formula", "Target"]
    for col_idx, header in enumerate(headers2, start=1):
        cell = ws.cell(row=14, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    metrics_data = [
        ("Overall Channel Coverage %", "=AVERAGE(C6:C11)", "≥85%"),
        ("Tier 1 Channels Avg (Critical)", "=AVERAGE(C6:C8)", "≥90%"),
        ("Tier 2 Channels Avg (High)", "=AVERAGE(C10:C11)", "≥80%"),
        ("Tier 3-4 Channels Avg (Med-Low)", "=C9", "≥70%"),
    ]
    
    for row_idx, metric_data in enumerate(metrics_data, start=15):
        for col_idx, value in enumerate(metric_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx == 1:
                cell.font = Font(bold=True)
            else:
                apply_style(cell, styles["normal_cell"])
            if col_idx == 2:
                cell.number_format = '0.0"%"'
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['B'].width = WIDTH_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A5'


def create_gap_analysis(ws, styles, validations):
    """Create Gap Analysis sheet."""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "CHANNEL PROTECTION GAP ANALYSIS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:L2')
    ws['A2'] = "Document all unprotected channels and remediation plans"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:L3')
    ws['A3'] = "Identify and prioritize gaps in channel protection coverage"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Gap ID", "Channel", "Gap Description", "Risk Level",
        "Exfiltration Risk", "Current Coverage %", "Target Coverage %",
        "Remediation Plan", "Owner", "Target Date", "Status", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated example gaps (gray rows)
    examples = [
        ("GAP-A812-3-001", "Endpoint", "Linux endpoints not protected by DLP agent", "High", "High", "0", "80", "Deploy endpoint DLP with Linux support or implement compensating controls", "IT Security Lead", "2025-06-30", "Not Started", "A812-3-GAP-001"),
        ("GAP-A812-3-002", "Mobile", "BYOD enrollment rate below target (<50%)", "Medium", "Medium", "50", "90", "Launch BYOD enrollment campaign with user training", "Mobile Admin", "2025-04-30", "In Progress", "A812-3-GAP-002"),
        ("GAP-A812-3-003", "Web", "Personal webmail access not blocked for high-risk users", "Critical", "Very High", "0", "100", "Deploy web filtering policy to block Gmail/Outlook.com for Confidential data users", "Network Security", "2025-03-31", "Planned", "A812-3-GAP-003"),
    ]
    
    for row_idx, example_data in enumerate(examples, start=6):
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
    
    # Blank input rows
    for row in range(9, 46):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
            # Auto-generate Gap ID
            if col == 1:
                cell.value = f"GAP-A812-3-{str(row-5).zfill(3)}"
    
    # Add data validations
    ws.add_data_validation(validations["channel_type"])
    for row in range(6, 46):
        validations["channel_type"].add(ws[f'B{row}'])
    
    ws.add_data_validation(validations["risk_level"])
    for row in range(6, 46):
        validations["risk_level"].add(ws[f'D{row}'])
    
    ws.add_data_validation(validations["exfiltration_risk"])
    for row in range(6, 46):
        validations["exfiltration_risk"].add(ws[f'E{row}'])
    
    ws.add_data_validation(validations["status"])
    for row in range(6, 46):
        validations["status"].add(ws[f'K{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['I'].width = WIDTH_WIDE
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    
    # Freeze panes
    ws.freeze_panes = 'A6'


# END OF PART 3

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 4 of 4)
# ============================================================================

def create_evidence_register(ws, styles, validations):
    """Create Evidence Register sheet."""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "EVIDENCE REGISTER"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Track all evidence collected for channel protection assessment"
    apply_style(ws['A2'], styles["subheader"])
    
    # Assessment question
    ws.merge_cells('A3:K3')
    ws['A3'] = "Document all evidence supporting channel coverage assessments"
    ws['A3'].font = Font(bold=True, size=11)
    
    # Column headers
    headers = [
        "Evidence ID", "Evidence Type", "Channel", "Description",
        "Collection Date", "Collected By", "Storage Location",
        "Verification Status", "Verified By", "Verification Date", "Notes"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Blank input rows (100 rows for evidence tracking)
    for row in range(6, 106):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["normal_cell"])
    
    # Add data validations
    ws.add_data_validation(validations["evidence_type"])
    for row in range(6, 106):
        validations["evidence_type"].add(ws[f'B{row}'])
    
    ws.add_data_validation(validations["channel_type"])
    for row in range(6, 106):
        validations["channel_type"].add(ws[f'C{row}'])
    
    ws.add_data_validation(validations["verification_status"])
    for row in range(6, 106):
        validations["verification_status"].add(ws[f'H{row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = WIDTH_WIDE
    ws.column_dimensions['G'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = WIDTH_WIDE
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = WIDTH_DESCRIPTION
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_summary_dashboard(ws, styles, validations):
    """Create Summary Dashboard sheet."""
    
    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "CHANNEL COVERAGE ASSESSMENT - EXECUTIVE SUMMARY"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35
    
    # Subheader
    ws.merge_cells('A2:D2')
    ws['A2'] = "ISO/IEC 27001:2022 - Control A.8.12: Data Leakage Prevention"
    apply_style(ws['A2'], styles["subheader"])
    
    # Section 1: Document Information
    ws['A4'] = "SECTION 1: DOCUMENT INFORMATION"
    ws['A4'].font = Font(bold=True, size=12)
    
    doc_info = [
        ("Assessment Date:", "[Input]"),
        ("Completed By:", "[Input]"),
        ("CISO Approval Status:", "[Dropdown]"),
        ("Last Review Date:", "[Input]"),
    ]
    
    row = 5
    for label, value in doc_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "Input" in value:
            apply_style(ws[f'B{row}'], styles["input_cell"])
        elif "Dropdown" in value:
            apply_style(ws[f'B{row}'], styles["input_cell"])
        row += 1
    
    # Add approval status validation
    ws.add_data_validation(validations["approval_status"])
    validations["approval_status"].add(ws['B7'])
    
    # Section 2: Key Performance Indicators
    row += 2
    ws[f'A{row}'] = "SECTION 2: KEY PERFORMANCE INDICATORS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    kpi_headers = ["KPI", "Formula/Value", "Target", "Status"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    kpis = [
        ("Overall Channel Coverage %", "=Coverage_Metrics!B15", "≥85%", "=IF(B13>=85%,\"\u2705\",IF(B13>=70%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Tier 1 Channels (Email/Web/Endpoint)", "=Coverage_Metrics!B16", "≥90%", "=IF(B14>=90%,\"\u2705\",IF(B14>=80%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Email Channel Coverage", "=Email_Channel!B25", "≥95%", "=IF(B15>=95%,\"\u2705\",IF(B15>=85%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Web/Cloud Channel Coverage", "=Web_Cloud_Channel!B25", "≥90%", "=IF(B16>=90%,\"\u2705\",IF(B16>=80%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Endpoint Channel Coverage", "=Endpoint_Channel!L30", "≥95%", "=IF(B17>=95%,\"\u2705\",IF(B17>=85%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Network Channel Coverage", "=Network_Channel!E20", "≥75%", "=IF(B18>=75%,\"\u2705\",IF(B18>=60%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Application Channel Coverage", "=Application_Channel!E20", "≥80%", "=IF(B19>=80%,\"\u2705\",IF(B19>=65%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Mobile Channel Coverage", "=Mobile_Channel!L20", "≥90%", "=IF(B20>=90%,\"\u2705\",IF(B20>=75%,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Total Gaps Identified", "=COUNTA(Gap_Analysis!A6:A45)-3", "≤10", "=IF(B21<=10,\"\u2705\",IF(B21<=15,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Critical/High Gaps", "=COUNTIFS(Gap_Analysis!D6:D45,\"Critical\")+COUNTIFS(Gap_Analysis!D6:D45,\"High\")", "0", "=IF(B22=0,\"\u2705\",IF(B22<=3,\"\u26A0\uFE0F\",\"\u274C\"))"),
    ]
    
    for kpi_name, formula, target, status_formula in kpis:
        ws[f'A{row}'] = kpi_name
        ws[f'B{row}'] = formula
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status_formula
        
        if "Coverage" in kpi_name:
            ws[f'B{row}'].number_format = '0.0"%"'
        
        apply_style(ws[f'A{row}'], styles["normal_cell"])
        apply_style(ws[f'B{row}'], styles["normal_cell"])
        apply_style(ws[f'C{row}'], styles["normal_cell"])
        apply_style(ws[f'D{row}'], styles["normal_cell"])
        row += 1
    
    # Domain 3 Compliance Score
    ws[f'A{row}'] = "Domain 3 Compliance Score"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'] = "=B13"
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'B{row}'].font = Font(bold=True, size=11)
    ws[f'C{row}'] = "≥85%"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = "=IF(B23>=85%,\"\u2705\",IF(B23>=70%,\"\u26A0\uFE0F\",\"\u274C\"))"
    ws[f'D{row}'].font = Font(bold=True, size=11)
    row += 1
    
    # Section 3: Gap Summary by Channel
    row += 2
    ws[f'A{row}'] = "SECTION 3: GAP SUMMARY BY CHANNEL"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    gap_headers = ["Channel", "Total Gaps", "Critical Gaps"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    channels_gap = [
        ("Email", "=COUNTIF(Gap_Analysis!B6:B45,\"Email\")", "=COUNTIFS(Gap_Analysis!B6:B45,\"Email\",Gap_Analysis!D6:D45,\"Critical\")"),
        ("Web/Cloud", "=COUNTIF(Gap_Analysis!B6:B45,\"Web\")", "=COUNTIFS(Gap_Analysis!B6:B45,\"Web\",Gap_Analysis!D6:D45,\"Critical\")"),
        ("Endpoint", "=COUNTIF(Gap_Analysis!B6:B45,\"Endpoint\")", "=COUNTIFS(Gap_Analysis!B6:B45,\"Endpoint\",Gap_Analysis!D6:D45,\"Critical\")"),
        ("Network", "=COUNTIF(Gap_Analysis!B6:B45,\"Network\")", "=COUNTIFS(Gap_Analysis!B6:B45,\"Network\",Gap_Analysis!D6:D45,\"Critical\")"),
        ("Application", "=COUNTIF(Gap_Analysis!B6:B45,\"App\")", "=COUNTIFS(Gap_Analysis!B6:B45,\"App\",Gap_Analysis!D6:D45,\"Critical\")"),
        ("Mobile", "=COUNTIF(Gap_Analysis!B6:B45,\"Mobile\")", "=COUNTIFS(Gap_Analysis!B6:B45,\"Mobile\",Gap_Analysis!D6:D45,\"Critical\")"),
    ]
    
    for channel_name, total_formula, critical_formula in channels_gap:
        ws[f'A{row}'] = channel_name
        ws[f'B{row}'] = total_formula
        ws[f'C{row}'] = critical_formula
        
        apply_style(ws[f'A{row}'], styles["normal_cell"])
        apply_style(ws[f'B{row}'], styles["normal_cell"])
        apply_style(ws[f'C{row}'], styles["normal_cell"])
        row += 1
    
    # Section 4: Evidence Completeness
    row += 2
    ws[f'A{row}'] = "SECTION 4: EVIDENCE COMPLETENESS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    evidence_metrics = [
        ("Evidence Items Collected", "=COUNTA(Evidence_Register!A6:A105)"),
        ("Evidence Items Verified", "=COUNTIF(Evidence_Register!H6:H105,\"Verified\")"),
        ("Evidence Completeness %", "=(B38/B37)*100"),
    ]
    
    for metric_name, metric_formula in evidence_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = metric_formula
        
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0"%"'
        
        apply_style(ws[f'B{row}'], styles["normal_cell"])
        row += 1
    
    # Section 5: Traffic Light Status
    row += 2
    ws[f'A{row}'] = "SECTION 5: OVERALL COMPLIANCE STATUS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "Overall Status:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=IF(B23>=85,\"\u2705 Compliant\",IF(B23>=70,\"\u26A0\uFE0F Partial Compliance\",\"\u274C Non-Compliant\"))"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    row += 1
    
    # Section 6: Approval Sign-Off
    row += 2
    ws[f'A{row}'] = "SECTION 6: APPROVAL SIGN-OFF"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    approval_headers = ["Approver", "Name", "Signature", "Date", "Status"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    approvers = [
        ("DLP Administrator", "[Input]", "[Input]", "=TODAY()", "[Dropdown]"),
        ("CISO", "[Input]", "[Input]", "=TODAY()", "[Dropdown]"),
        ("DPO", "[Input]", "[Input]", "=TODAY()", "[Dropdown]"),
    ]
    
    for approver_role, name, signature, date, status in approvers:
        ws[f'A{row}'] = approver_role
        ws[f'B{row}'] = name
        ws[f'C{row}'] = signature
        ws[f'D{row}'] = date
        ws[f'E{row}'] = status
        
        apply_style(ws[f'A{row}'], styles["normal_cell"])
        
        if "Input" in name:
            apply_style(ws[f'B{row}'], styles["input_cell"])
        if "Input" in signature:
            apply_style(ws[f'C{row}'], styles["input_cell"])
        
        apply_style(ws[f'D{row}'], styles["normal_cell"])
        
        if "Dropdown" in status:
            apply_style(ws[f'E{row}'], styles["input_cell"])
        
        row += 1
    
    # Add approval status validations
    for approval_row in range(row-3, row):
        ws.add_data_validation(validations["approval_status"])
        validations["approval_status"].add(ws[f'E{approval_row}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A5'


# ============================================================================
# SECTION 5: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the Channel Coverage Assessment workbook."""
    
    logger.info("=" * 78)
    logger.info(f"{WORKBOOK_ID} - {ASSESSMENT_AREA} Generator")
    logger.info(f"ISO/IEC 27001:2022 Control {CONTROL_ID}")
    logger.info("=" * 78)
    
    wb = create_workbook()
    styles = setup_styles()
    validations = create_data_validations()
    
    logger.info("\n[1/12] Creating Instructions & Legend...")
    create_instructions(wb["Instructions_Legend"], styles)
    
    logger.info("[2/12] Creating Channel Overview...")
    create_channel_overview(wb["Channel_Overview"], styles, validations)
    
    logger.info("[3/12] Creating Email Channel...")
    create_email_channel(wb["Email_Channel"], styles, validations)
    
    logger.info("[4/12] Creating Web/Cloud Channel...")
    create_web_cloud_channel(wb["Web_Cloud_Channel"], styles, validations)
    
    logger.info("[5/12] Creating Endpoint Channel...")
    create_endpoint_channel(wb["Endpoint_Channel"], styles, validations)
    
    logger.info("[6/12] Creating Network Channel...")
    create_network_channel(wb["Network_Channel"], styles, validations)
    
    logger.info("[7/12] Creating Application Channel...")
    create_application_channel(wb["Application_Channel"], styles, validations)
    
    logger.info("[8/12] Creating Mobile Channel...")
    create_mobile_channel(wb["Mobile_Channel"], styles, validations)
    
    logger.info("[9/12] Creating Coverage Metrics...")
    create_coverage_metrics(wb["Coverage_Metrics"], styles)
    
    logger.info("[10/12] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap_Analysis"], styles, validations)
    
    logger.info("[11/12] Creating Evidence Register...")
    create_evidence_register(wb["Evidence_Register"], styles, validations)
    
    logger.info("[12/12] Creating Summary Dashboard...")
    create_summary_dashboard(wb["Summary_Dashboard"], styles, validations)
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.12.3_Channel_Coverage_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  \u2022 Instructions & Legend")
    logger.info("  \u2022 Channel Overview (6 channels summary)")
    logger.info("  \u2022 Email Channel (20 assessment rows)")
    logger.info("  \u2022 Web/Cloud Channel (20 assessment rows)")
    logger.info("  \u2022 Endpoint Channel (25 assessment rows)")
    logger.info("  \u2022 Network Channel (15 assessment rows)")
    logger.info("  \u2022 Application Channel (15 assessment rows)")
    logger.info("  \u2022 Mobile Channel (15 assessment rows)")
    logger.info("  \u2022 Coverage Metrics (weighted coverage calculations)")
    logger.info("  \u2022 Gap Analysis (40 gap rows)")
    logger.info("  \u2022 Evidence Register (100 evidence rows)")
    logger.info("  \u2022 Summary Dashboard (KPIs + compliance scoring)")
    logger.info("\nTotal Assessment Items: ~90 channel protection checkpoints")
    logger.info("\nChannel Protection Tiers:")
    logger.info("  Tier 1 (Critical): Email, Web, USB - Months 1-3")
    logger.info("  Tier 2 (High): Cloud, Mobile - Months 4-6")
    logger.info("  Tier 3 (Medium): Network shares, Print - Months 7-9")
    logger.info("  Tier 4 (Low): Bluetooth, Optical drives - Months 10-12")
    logger.info("\n" + "=" * 78)
    logger.info("\n🎯 Remember: The best DLP policy is worthless if you forget")
    logger.info("   to protect the USB port. — Murphy's Law of Data Leakage")
    logger.info("\n" + "=" * 78)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
