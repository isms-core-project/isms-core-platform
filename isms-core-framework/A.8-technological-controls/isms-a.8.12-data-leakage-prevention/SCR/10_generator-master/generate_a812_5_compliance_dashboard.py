#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.5 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Master Consolidation Dashboard (Aggregates Domains 1-4)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific reporting requirements, KPI definitions, and
governance structure.

Key customization areas:
1. External workbook paths (update to your actual file locations)
2. KPI definitions and thresholds (per your governance requirements)
3. Risk scoring criteria (aligned with your risk framework)
4. Compliance mapping (specific to your regulatory obligations)
5. Approval workflow (based on your organizational structure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates executive compliance dashboard that consolidates assessment data from
all four DLP evaluation domains into unified view for CISO reporting and audit
evidence presentation.

This dashboard uses EXTERNAL WORKBOOK FORMULAS to pull data from normalized
assessment files, enabling real-time compliance status monitoring and gap
tracking across the complete DLP control framework.

The dashboard provides:
• Executive summary with overall DLP compliance score
• Consolidated gap analysis from all 4 assessment domains
• Risk register for DLP-related information security risks
• Remediation roadmap with timelines and ownership
• Evidence master index consolidating all audit evidence
• KPI dashboard tracking key performance indicators
• Compliance scorecard mapping to ISO 27001 requirements
• Management approval workflow

--------------------------------------------------------------------------------
DASHBOARD STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.5_Compliance_Dashboard_YYYYMMDD.xlsx

Sheets (9 total):
1. Executive_Summary - High-level metrics and compliance scores
2. Consolidated_Gap_Analysis - All gaps from domains 1-4
3. Risk_Register - DLP-related risks and treatment plans
4. Remediation_Roadmap - Action plans with timelines and owners
5. Evidence_Master_Index - Consolidated evidence from all domains
6. KPI_Dashboard - Performance indicators and trends
7. Compliance_Scorecard - ISO 27001 A.8.12 requirements mapping
8. Approval_Sign_Off - CISO and management approvals
9. Document_Control - Version history and change log

--------------------------------------------------------------------------------
CRITICAL PREREQUISITE - FILE NORMALIZATION
--------------------------------------------------------------------------------

⚠️  IMPORTANT: This dashboard uses external workbook formulas that require
source assessment files to have EXACT normalized filenames.

BEFORE generating this dashboard, you MUST run:
    python3 normalize_assessment_files_a812.py

This creates normalized files:
    ISMS-IMP-A.8.12.1.xlsx (Infrastructure Assessment)
    ISMS-IMP-A.8.12.2.xlsx (Data Classification Assessment)
    ISMS-IMP-A.8.12.3.xlsx (Channel Coverage Assessment)
    ISMS-IMP-A.8.12.4.xlsx (Monitoring & Response Assessment)

All 5 files (4 assessments + dashboard) must be in the SAME DIRECTORY for
external formulas to work correctly.

If you skip normalization or place files in different directories, external
formulas will show #REF! errors and dashboard will not function.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Microsoft Excel or LibreOffice Calc (for opening dashboard)

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Complete Workflow:

1. Generate all 4 assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py
       python3 generate_a812_2_data_classification.py
       python3 generate_a812_3_channel_coverage.py
       python3 generate_a812_4_monitoring_response.py

2. Complete assessments (manual work by security team, data owners, SOC)

3. Normalize assessment filenames (CRITICAL STEP):
       python3 normalize_assessment_files_a812.py
   
   This step creates exact filenames for external formulas.
   Without normalization, external formulas will fail.

4. Generate this dashboard:
       python3 generate_a812_5_compliance_dashboard.py    ← YOU ARE HERE

5. Place ALL 5 files in same directory (e.g., Dashboard_Sources/)
   Files must be co-located in the same folder.

6. Open dashboard in Excel:
   - Click "Enable Content" (to allow formulas)
   - Click "Update Links" (to refresh data from source workbooks)

7. (Optional) Run consolidation for static snapshot:
       python3 consolidate_a812_dashboard.py [dashboard_filename]
   
   This creates point-in-time snapshot useful for archival/audit packages.

--------------------------------------------------------------------------------
EXTERNAL FORMULAS vs DATA CONSOLIDATION
--------------------------------------------------------------------------------

This dashboard supports TWO operational modes:

**Mode 1: External Formulas (Default, Recommended)**
Dashboard contains formulas like:
    ='[ISMS-IMP-A.8.12.1.xlsx]Gap_Analysis'!$B$15

Characteristics:
• Data updates AUTOMATICALLY when source assessment files change
• Requires all files in same directory (co-located)
• Real-time compliance status monitoring
• Click "Update Links" when opening dashboard
• Best for: Day-to-day DLP compliance management

Benefits:
• No manual data entry required
• Always current compliance status
• Changes in assessments immediately reflected
• Reduces maintenance effort

Limitations:
• Files must be co-located (same directory)
• Excel security warnings on open (normal behavior)
• Network drives may require trusted locations configuration

**Mode 2: Static Data Consolidation (Alternative)**
Run consolidation script:
    python3 consolidate_a812_dashboard.py [dashboard_file]

Characteristics:
• Copies actual data values (not formulas) into dashboard
• Creates point-in-time compliance snapshot
• Self-contained workbook (no external dependencies)
• Must re-run script when source data changes
• Best for: Quarterly audit packages, Board presentations

Benefits:
• Single file distribution (no dependencies)
• Historical compliance tracking (create dated snapshots)
• No Excel security warnings
• Works when files in different locations

Limitations:
• Manual refresh required (re-run script)
• Not real-time (snapshot only)
• Larger file size (data duplicated)

**Recommended Approach:**
Use Mode 1 (external formulas) for day-to-day management.
Use Mode 2 (consolidation) for quarterly audit evidence packages.

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Master consolidation layer - aggregates all 4 assessment domains
    Final output for executive reporting and audit evidence
    
Related Documents:
    Policy:         ISMS-POL-A.8.12 (Complete DLP Policy)
    Implementation: ISMS-IMP-A.8.12 (Implementation Guide)
    Assessments:    ISMS-IMP-A.8.12.1 through A.8.12.4

Data Sources:
    Domain 1: ISMS-IMP-A.8.12.1.xlsx (DLP Infrastructure)
    Domain 2: ISMS-IMP-A.8.12.2.xlsx (Data Classification)
    Domain 3: ISMS-IMP-A.8.12.3.xlsx (Channel Coverage)
    Domain 4: ISMS-IMP-A.8.12.4.xlsx (Monitoring & Response)

Output Consumers:
    • CISO (executive compliance reporting)
    • ISO 27001 auditors (control A.8.12 evidence)
    • Board/executive management (risk oversight)
    • Security team (gap tracking and remediation)
    • Data Protection Officer (regulatory compliance)

Complete Workflow:
    1. Generate 4 assessment workbooks (scripts 1-4)
    2. Complete assessments (manual - various teams)
    3. Normalize filenames (normalize_assessment_files_a812.py)
    4. Generate dashboard (THIS SCRIPT)
    5. Open dashboard and enable links
    6. (Optional) Consolidate for static snapshot
    7. Present to CISO/auditors
    8. Track remediation actions
    9. Quarterly reassessment cycle

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Dashboard Type:       Master Consolidation (Aggregates Domains 1-4)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organization] ISMS Implementation Team
License:              [Organization License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - ISO/IEC 27004:2016 (Information Security Monitoring and Measurement)
    - ISO/IEC 27005:2022 (Information Security Risk Management)
    - Swiss FADP (Federal Act on Data Protection)
    - EU GDPR (General Data Protection Regulation)
    - NIST SP 800-53 (Security and Privacy Controls)
    - CIS Controls v8 (Center for Internet Security)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

File Location Requirements:
    All 5 workbooks MUST be in the same directory:
    • ISMS-IMP-A.8.12.1.xlsx
    • ISMS-IMP-A.8.12.2.xlsx
    • ISMS-IMP-A.8.12.3.xlsx
    • ISMS-IMP-A.8.12.4.xlsx
    • ISMS-IMP-A.8.12.5_Compliance_Dashboard_YYYYMMDD.xlsx
    
    External formulas will break (#REF! errors) if files are in different directories.
    
    Recommended directory structure:
    /path/to/DLP_Compliance/
        ├── ISMS-IMP-A.8.12.1.xlsx
        ├── ISMS-IMP-A.8.12.2.xlsx
        ├── ISMS-IMP-A.8.12.3.xlsx
        ├── ISMS-IMP-A.8.12.4.xlsx
        └── ISMS-IMP-A.8.12.5_Compliance_Dashboard_20250125.xlsx

Excel Security Settings:
    When opening dashboard, you will see security warnings:
    
    1. "Security Warning: External Data Connections have been disabled"
       → Click "Enable Content" button
       → This is normal - Excel blocks external formulas by default
    
    2. "This workbook contains links to other data sources"
       → Click "Update" button to refresh data
       → Click "Don't Update" if you want to keep old data (not recommended)
    
    These security warnings are EXPECTED BEHAVIOR for workbooks with external
    formulas. They do not indicate a security problem.
    
    If your organization's Excel security policy prevents external formulas:
    • Use consolidation script instead (Mode 2)
    • Or configure trusted locations in Excel settings

Network Drives and SharePoint:
    External formulas work on network drives but may require configuration:
    • Add network path to Excel Trusted Locations
    • Use UNC paths (\\\\server\\share) not mapped drives (Z:\\)
    • SharePoint/OneDrive: May require desktop sync client
    
    If external formulas don't work on network drives:
    • Use consolidation script (Mode 2)
    • Or work with local file copies

Audit Evidence Requirements:
    This dashboard serves as primary ISO 27001:2022 Control A.8.12 audit evidence.
    
    For audit readiness:
    • Retain completed dashboard with source assessment workbooks
    • Create quarterly snapshots (use consolidation script)
    • Document evidence in Evidence_Master_Index sheet
    • Obtain CISO approval (Approval_Sign_Off sheet)
    • Maintain version history (Document_Control sheet)
    
    Retention requirements:
    • Active compliance: Current dashboard + previous quarter
    • Audit cycle: 3 years of quarterly snapshots
    • Regulatory: Per Swiss FADP/EU GDPR requirements (typically 3-7 years)
    
    Classification: [Organization] Internal/Confidential
    
    Sensitive information in dashboard:
    • Security gaps and vulnerabilities
    • Risk assessments and treatment plans
    • Compliance scores and deficiencies
    • Audit evidence references

Review and Update Cycle:
    Monthly:
        • Review KPI_Dashboard for performance trends
        • Track remediation progress (Remediation_Roadmap)
        • Update gap status as remediation completes
    
    Quarterly:
        • Complete reassessment of all 4 domains
        • Update source assessment workbooks
        • Refresh dashboard (click "Update Links")
        • Create static snapshot for audit (consolidation script)
        • Present to CISO and management
    
    Annually:
        • Full DLP framework reassessment
        • Update policies if requirements changed
        • Review and update data classification schema
        • Assess new DLP technologies
        • Complete audit cycle evidence package

Performance Considerations:
    Dashboard typically opens in 5-10 seconds with external formulas.
    
    If performance issues occur:
    • Check all source files are closed (don't open simultaneously)
    • Verify files are local (not slow network drives)
    • Consider using consolidation script for large datasets
    • Close other Excel workbooks to free memory
    
    File sizes:
    • Dashboard: ~2-3 MB (with formulas)
    • Consolidated dashboard: ~4-5 MB (with data values)
    • Total framework: ~15-20 MB (all 5 workbooks)

Troubleshooting Common Issues:
    Problem: #REF! errors in dashboard
    Solution: Files not in same directory - run normalization and move files
    
    Problem: "Cannot update link" error
    Solution: Source assessment files missing - check all 4 workbooks present
    
    Problem: Old data showing despite changes in source
    Solution: Click "Update Links" when opening, or close and reopen dashboard
    
    Problem: Excel security prevents external formulas
    Solution: Use consolidation script (Mode 2) or configure trusted locations
    
    Problem: SharePoint/OneDrive sync issues
    Solution: Sync files locally, or use consolidation script

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.5"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: CONSTANTS & CONFIGURATION
# ============================================================================

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.5"
RELATED_POLICY = "ISMS-POL-A.8.12 (All DLP Policies - Master)"
ASSESSMENT_AREA = "Compliance Dashboard - Master Consolidation"

# Normalized workbook filenames (created by normalize_assessment_files_a812.py)
WORKBOOK_DOMAIN_1 = "ISMS-IMP-A.8.12.1.xlsx"
WORKBOOK_DOMAIN_2 = "ISMS-IMP-A.8.12.2.xlsx"
WORKBOOK_DOMAIN_3 = "ISMS-IMP-A.8.12.3.xlsx"
WORKBOOK_DOMAIN_4 = "ISMS-IMP-A.8.12.4.xlsx"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "E7E6E6"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "B4C7E7"         # Light blue (\u1F4CB Planned)
COLOR_WARNING = "FF6B6B"         # Red (for critical warnings)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 40


# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    sheets = [
        "Instructions_Legend",
        "Executive_Summary",
        "Domain_Rollup_Summary",
        "Consolidated_Gap_Analysis",
        "Risk_Register",
        "Remediation_Roadmap",
        "Evidence_Master_Index",
        "Trend_Analysis",
        "KPI_Dashboard",
        "Budget_Planning",
        "CISO_DPO_Approval",
        "Summary_Dashboard",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    """
    return {
        "header": {
            "font": Font(name="Arial", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Arial", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "warning": {
            "font": Font(name="Arial", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )
        },
        "column_header": {
            "font": Font(name="Arial", size=11, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "font": Font(name="Arial", size=10),
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "font": Font(name="Arial", size=10, italic=True),
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "data_cell": {
            "font": Font(name="Arial", size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
    }


def apply_style(cell, style_dict):
    """
    Apply style template to cell by creating NEW instances.
    """
    if "font" in style_dict:
        f = style_dict["font"]
        cell.font = Font(
            name=f.name,
            size=f.size,
            bold=f.bold,
            italic=f.italic,
            color=f.color
        )
    
    if "fill" in style_dict:
        f = style_dict["fill"]
        cell.fill = PatternFill(
            start_color=f.start_color.rgb if f.start_color else None,
            end_color=f.end_color.rgb if f.end_color else None,
            fill_type=f.fill_type
        )
    
    if "alignment" in style_dict:
        a = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=a.horizontal,
            vertical=a.vertical,
            wrap_text=a.wrap_text
        )
    
    if "border" in style_dict:
        b = style_dict["border"]
        cell.border = Border(
            left=Side(style=b.left.style) if b.left else None,
            right=Side(style=b.right.style) if b.right else None,
            top=Side(style=b.top.style) if b.top else None,
            bottom=Side(style=b.bottom.style) if b.bottom else None
        )


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def setup_data_validations(wb):
    """Setup all data validation dropdowns used in the workbook."""
    
    # Source domains
    val_domain = DataValidation(
        type="list",
        formula1='"Domain 1,Domain 2,Domain 3,Domain 4,Cross-Domain"',
        allow_blank=False
    )
    
    # Risk levels
    val_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    
    # Risk categories
    val_risk_category = DataValidation(
        type="list",
        formula1='"Insider Threat,External Attack,Compliance,Operational,Technical"',
        allow_blank=False
    )
    
    # Likelihood
    val_likelihood = DataValidation(
        type="list",
        formula1='"Very High,High,Medium,Low,Very Low"',
        allow_blank=False
    )
    
    # Evidence types
    val_evidence = DataValidation(
        type="list",
        formula1='"Config,Screenshot,Log,Report,Dashboard,Other"',
        allow_blank=False
    )
    
    # Verification status
    val_verification = DataValidation(
        type="list",
        formula1='"Verified,Pending,Rejected"',
        allow_blank=False
    )
    
    # Gap status
    val_gap_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,Blocked"',
        allow_blank=False
    )
    
    return {
        "domain": val_domain,
        "risk": val_risk,
        "risk_category": val_risk_category,
        "likelihood": val_likelihood,
        "evidence": val_evidence,
        "verification": val_verification,
        "gap_status": val_gap_status,
    }


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 1
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create Instructions_Legend sheet with normalization workflow instructions."""
    ws = wb["Instructions_Legend"]
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{WORKBOOK_ID} - {ASSESSMENT_AREA}"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # SETUP INSTRUCTIONS BOX (rows 3-20)
    row = 3
    ws.merge_cells(f'A{row}:F{row+17}')
    setup_text = """⚙️ DASHBOARD SETUP INSTRUCTIONS ⚙️

This workbook consolidates data from Domains 1-4 using external formulas.

REQUIRED SETUP STEPS:

Step 1: Complete all 4 assessment workbooks (Domains 1-4)
Step 2: Run the normalization script:
        python3 normalize_assessment_files_a812.py
        
        This creates standardized filenames in Dashboard_Sources folder:
        - ISMS-IMP-A.8.12.1.xlsx (DLP Infrastructure)
        - ISMS-IMP-A.8.12.2.xlsx (Data Classification)
        - ISMS-IMP-A.8.12.3.xlsx (Channel Coverage)
        - ISMS-IMP-A.8.12.4.xlsx (Monitoring & Response)

Step 3: Place THIS dashboard file in the Dashboard_Sources folder
        (same folder as normalized assessment workbooks)

Step 4: Open this dashboard
Step 5: Click "Update Links" when prompted
Step 6: Dashboard auto-populates with compliance data!

\u2705 External formulas automatically reference normalized workbooks
\u2705 No manual Find & Replace required!"""
    
    ws[f'A{row}'] = setup_text
    ws[f'A{row}'].font = Font(name="Arial", size=11, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[f'A{row}'].border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )
    ws.row_dimensions[row].height = 320
    
    row = 21
    
    # Document metadata
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "DOCUMENT METADATA"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    metadata = [
        ("Document ID:", WORKBOOK_ID),
        ("Assessment Area:", ASSESSMENT_AREA),
        ("Related Domains:", "1, 2, 3, 4"),
        ("Version:", WORKBOOK_VERSION),
        ("ISO Control:", CONTROL_ID),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Organization metadata
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ORGANIZATION METADATA (Complete the yellow fields)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    org_fields = [
        "Assessment Date:",
        "Completed By (Name):",
        "Completed By (Role):",
        "Organization Name:",
        "Review Cycle:",
    ]
    
    for field in org_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(bold=True)
        apply_style(ws[f'B{row}'], styles["input_cell"])
        if field == "Review Cycle:":
            ws[f'B{row}'] = "Quarterly"
        row += 1
    
    # How to use
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    instructions = [
        "1. Review Executive_Summary (CISO one-page dashboard)",
        "2. Check Domain_Rollup_Summary (individual domain scores)",
        "3. Review Consolidated_Gap_Analysis (auto-populated from domains)",
        "4. Review Risk_Register (DLP-specific risks)",
        "5. Create Remediation_Roadmap (prioritized actions)",
        "6. Track trends in Trend_Analysis (quarterly progress)",
        "7. Review Budget_Planning (costs and staffing)",
        "8. Obtain final sign-off in CISO_DPO_Approval sheet",
        "9. Review Summary_Dashboard for master compliance score",
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    # Legend
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "LEGEND - COMPLIANCE SCORING"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    legend_items = [
        ("\u2705 Pass", "Overall DLP Compliance ≥ 85%"),
        ("\u26A0\uFE0F Conditional Pass", "70-84% (gaps must be remediated within 90 days)"),
        ("\u274C Fail", "<70% (significant remediation required)"),
    ]
    
    for symbol, description in legend_items:
        ws[f'A{row}'] = symbol
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = description
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_executive_summary_sheet(wb, styles):
    """Create Executive_Summary sheet (one-page CISO dashboard)."""
    ws = wb["Executive_Summary"]
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = "EXECUTIVE SUMMARY - DLP Program Compliance Dashboard"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Metadata
    row = 3
    ws[f'A{row}'] = "Report Date:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=TODAY()"
    ws[f'B{row}'].number_format = 'DD.MM.YYYY'
    
    row += 1
    ws[f'A{row}'] = "Reporting Period:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 1
    ws[f'A{row}'] = "Prepared By:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 2
    
    # Overall compliance summary section
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "OVERALL DLP COMPLIANCE SUMMARY"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    # Domain scores with external formulas
    headers = ["Domain", "Compliance Score", "Target", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    domains = [
        ("Domain 1 - Infrastructure", f"='[{WORKBOOK_DOMAIN_1}]Summary_Dashboard'!$B$19", "85%"),
        ("Domain 2 - Classification", f"='[{WORKBOOK_DOMAIN_2}]Summary_Dashboard'!$B$18", "85%"),
        ("Domain 3 - Channels", f"='[{WORKBOOK_DOMAIN_3}]Summary_Dashboard'!$B$23", "85%"),
        ("Domain 4 - Monitoring", f"='[{WORKBOOK_DOMAIN_4}]Summary_Dashboard'!$B$17", "85%"),
    ]
    
    domain_start_row = row
    for domain_name, formula, target in domains:
        ws[f'A{row}'] = domain_name
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0"%"'
        ws[f'C{row}'] = target
        ws[f'D{row}'] = f'=IF(B{row}>=85,"\u2705 Pass",IF(B{row}>=70,"\u26A0\uFE0F Conditional","\u274C Fail"))'
        row += 1
    
    # Overall weighted compliance
    ws[f'A{row}'] = "OVERALL DLP COMPLIANCE"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'B{row}'] = f'=(B{domain_start_row}*0.25+B{domain_start_row+1}*0.20+B{domain_start_row+2}*0.30+B{domain_start_row+3}*0.25)'
    ws[f'B{row}'].font = Font(bold=True, size=14)
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'C{row}'] = ">85%"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=IF(B{row}>=85,"\u2705 Pass",IF(B{row}>=70,"\u26A0\uFE0F Conditional","\u274C Fail"))'
    ws[f'D{row}'].font = Font(bold=True, size=14)
    
    row += 2
    
    # Key metrics section
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "KEY METRICS BY DOMAIN"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    metrics = [
        ("DOMAIN 1 - DLP INFRASTRUCTURE", ""),
        ("  Total DLP technologies deployed", f"='[{WORKBOOK_DOMAIN_1}]Summary_Dashboard'!$B$5"),
        ("  Endpoint coverage %", f"='[{WORKBOOK_DOMAIN_1}]Summary_Dashboard'!$B$6"),
        ("  Email DLP coverage %", f"='[{WORKBOOK_DOMAIN_1}]Summary_Dashboard'!$B$7"),
        ("", ""),
        ("DOMAIN 2 - DATA CLASSIFICATION", ""),
        ("  Data classification completeness %", f"='[{WORKBOOK_DOMAIN_2}]Summary_Dashboard'!$B$5"),
        ("  Sensitive data categories inventoried", f"='[{WORKBOOK_DOMAIN_2}]Summary_Dashboard'!$B$6"),
        ("  Data owners assigned %", f"='[{WORKBOOK_DOMAIN_2}]Summary_Dashboard'!$B$7"),
        ("", ""),
        ("DOMAIN 3 - CHANNEL COVERAGE", ""),
        ("  Overall channel coverage %", f"='[{WORKBOOK_DOMAIN_3}]Summary_Dashboard'!$B$5"),
        ("  Email channel coverage %", f"='[{WORKBOOK_DOMAIN_3}]Summary_Dashboard'!$B$8"),
        ("  Unprotected channels count", f"='[{WORKBOOK_DOMAIN_3}]Summary_Dashboard'!$B$17"),
        ("", ""),
        ("DOMAIN 4 - MONITORING & RESPONSE", ""),
        ("  False positive rate %", f"='[{WORKBOOK_DOMAIN_4}]Summary_Dashboard'!$B$11"),
        ("  SLA compliance %", f"='[{WORKBOOK_DOMAIN_4}]Summary_Dashboard'!$B$12"),
        ("  Alert triage rate %", f"='[{WORKBOOK_DOMAIN_4}]Summary_Dashboard'!$B$13"),
    ]
    
    for metric_name, formula in metrics:
        if metric_name == "":
            # Blank row
            row += 1
            continue
        
        ws[f'A{row}'] = metric_name
        if metric_name.startswith("DOMAIN"):
            ws[f'A{row}'].font = Font(bold=True, size=11)
        
        if formula:
            ws[f'B{row}'] = formula
            if "%" in metric_name:
                ws[f'B{row}'].number_format = '0.0"%"'
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_WIDE
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_domain_rollup_summary_sheet(wb, styles):
    """Create Domain_Rollup_Summary sheet."""
    ws = wb["Domain_Rollup_Summary"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "Domain Rollup Summary - Detailed Compliance View"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Column headers
    headers = ["Domain", "Compliance Score", "Critical Gaps", "High Gaps", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Domain 1
    ws[f'A{row}'] = "Domain 1 - DLP Infrastructure"
    ws[f'B{row}'] = f"='[{WORKBOOK_DOMAIN_1}]Summary_Dashboard'!$B$19"
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'C{row}'] = f"='[{WORKBOOK_DOMAIN_1}]Gap_Analysis'!$B$3"
    ws[f'D{row}'] = f"='[{WORKBOOK_DOMAIN_1}]Gap_Analysis'!$B$4"
    ws[f'E{row}'] = f'=IF(B{row}>=85,"\u2705",IF(B{row}>=70,"\u26A0\uFE0F","\u274C"))'
    row += 1
    
    # Domain 2
    ws[f'A{row}'] = "Domain 2 - Data Classification"
    ws[f'B{row}'] = f"='[{WORKBOOK_DOMAIN_2}]Summary_Dashboard'!$B$18"
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'C{row}'] = f"='[{WORKBOOK_DOMAIN_2}]Gap_Analysis'!$B$3"
    ws[f'D{row}'] = f"='[{WORKBOOK_DOMAIN_2}]Gap_Analysis'!$B$4"
    ws[f'E{row}'] = f'=IF(B{row}>=85,"\u2705",IF(B{row}>=70,"\u26A0\uFE0F","\u274C"))'
    row += 1
    
    # Domain 3
    ws[f'A{row}'] = "Domain 3 - Channel Coverage"
    ws[f'B{row}'] = f"='[{WORKBOOK_DOMAIN_3}]Summary_Dashboard'!$B$23"
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'C{row}'] = f"='[{WORKBOOK_DOMAIN_3}]Gap_Analysis'!$B$3"
    ws[f'D{row}'] = f"='[{WORKBOOK_DOMAIN_3}]Gap_Analysis'!$B$4"
    ws[f'E{row}'] = f'=IF(B{row}>=85,"\u2705",IF(B{row}>=70,"\u26A0\uFE0F","\u274C"))'
    row += 1
    
    # Domain 4
    ws[f'A{row}'] = "Domain 4 - Monitoring & Response"
    ws[f'B{row}'] = f"='[{WORKBOOK_DOMAIN_4}]Summary_Dashboard'!$B$17"
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'C{row}'] = f"='[{WORKBOOK_DOMAIN_4}]Gap_Analysis'!$B$3"
    ws[f'D{row}'] = f"='[{WORKBOOK_DOMAIN_4}]Gap_Analysis'!$B$4"
    ws[f'E{row}'] = f'=IF(B{row}>=85,"\u2705",IF(B{row}>=70,"\u26A0\uFE0F","\u274C"))'
    row += 1
    
    row += 1
    
    # Additional info section
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "DOMAIN DETAILS"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    details = [
        ("Domain 1: DLP Infrastructure", "Assesses DLP technology deployment and coverage"),
        ("Domain 2: Data Classification", "Evaluates data classification and ownership"),
        ("Domain 3: Channel Coverage", "Measures DLP protection across communication channels"),
        ("Domain 4: Monitoring & Response", "Reviews incident detection and response capabilities"),
    ]
    
    for domain, description in details:
        ws[f'A{row}'] = domain
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        ws.merge_cells(f'B{row}:E{row}')
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_consolidated_gap_analysis_sheet(wb, styles, validations):
    """Create Consolidated_Gap_Analysis sheet."""
    ws = wb["Consolidated_Gap_Analysis"]
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = "Consolidated Gap Analysis - All Domains"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Instructions
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "\u1F4CB Copy gaps from individual domain workbooks to consolidate here for enterprise-wide visibility"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
    
    row += 2
    
    # Column headers
    headers = ["Domain", "Gap ID", "Gap Description", "Severity", "Current State", "Target State", "Owner", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Add 100 rows for gap entries
    for i in range(100):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Add data validation for Domain column
        validations["domain"].add(ws.cell(row=row, column=1))
        
        # Add data validation for Severity column
        validations["risk"].add(ws.cell(row=row, column=4))
        
        # Add data validation for Status column
        validations["gap_status"].add(ws.cell(row=row, column=8))
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_WIDE
    ws.column_dimensions['F'].width = WIDTH_WIDE
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_risk_register_sheet(wb, styles, validations):
    """Create Risk_Register sheet."""
    ws = wb["Risk_Register"]
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "DLP Risk Register"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Column headers
    headers = ["Risk ID", "Domain", "Risk Description", "Category", "Likelihood", "Impact", "Risk Score", "Inherent Risk", "Residual Risk", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Pre-populated example risks
    example_risks = [
        ("R-001", "Domain 1", "Insufficient endpoint DLP coverage", "Technical", "High", "High", "=E4&\"-\"&F4", "High", "Medium", "Active"),
        ("R-002", "Domain 2", "Incomplete data classification", "Compliance", "Medium", "High", "=E5&\"-\"&F5", "High", "Low", "Mitigated"),
        ("R-003", "Domain 3", "Email channel unprotected", "Technical", "High", "Critical", "=E6&\"-\"&F6", "Critical", "High", "Active"),
        ("R-004", "Domain 4", "High false positive rate", "Operational", "Medium", "Medium", "=E7&\"-\"&F7", "Medium", "Low", "In Progress"),
        ("R-005", "Cross-Domain", "Lack of DLP policy awareness", "Compliance", "High", "Medium", "=E8&\"-\"&F8", "High", "Medium", "Active"),
    ]
    
    for risk_data in example_risks:
        for col_idx, value in enumerate(risk_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # Add 35 more rows for custom entries
    for i in range(35):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Risk score formula
        ws.cell(row=row, column=7).value = f"=E{row}&\"-\"&F{row}"
        
        # Add data validations
        validations["domain"].add(ws.cell(row=row, column=2))
        validations["risk_category"].add(ws.cell(row=row, column=4))
        validations["likelihood"].add(ws.cell(row=row, column=5))
        validations["likelihood"].add(ws.cell(row=row, column=6))
        validations["risk"].add(ws.cell(row=row, column=8))
        validations["risk"].add(ws.cell(row=row, column=9))
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_MEDIUM
    ws.column_dimensions['I'].width = WIDTH_MEDIUM
    ws.column_dimensions['J'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_remediation_roadmap_sheet(wb, styles, validations):
    """Create Remediation_Roadmap sheet."""
    ws = wb["Remediation_Roadmap"]
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "Remediation Roadmap"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Column headers
    headers = ["Action ID", "Domain", "Action Description", "Priority", "Owner", "Start Date", "Target Date", "Actual Date", "Budget", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Add 50 rows for remediation actions
    for i in range(50):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Add data validations
        validations["domain"].add(ws.cell(row=row, column=2))
        validations["risk"].add(ws.cell(row=row, column=4))
        validations["gap_status"].add(ws.cell(row=row, column=10))
        
        # Date formatting
        ws.cell(row=row, column=6).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=7).number_format = 'DD.MM.YYYY'
        ws.cell(row=row, column=8).number_format = 'DD.MM.YYYY'
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_MEDIUM
    ws.column_dimensions['I'].width = WIDTH_MEDIUM
    ws.column_dimensions['J'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_evidence_master_index_sheet(wb, styles, validations):
    """Create Evidence_Master_Index sheet."""
    ws = wb["Evidence_Master_Index"]
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = "Evidence Master Index"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Column headers
    headers = ["Evidence ID", "Domain", "Control/Requirement", "Evidence Type", "File Location", "Collected Date", "Verified By", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Add 100 rows for evidence entries
    for i in range(100):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["input_cell"])
        
        # Add data validations
        validations["domain"].add(ws.cell(row=row, column=2))
        validations["evidence"].add(ws.cell(row=row, column=4))
        validations["verification"].add(ws.cell(row=row, column=8))
        
        # Date formatting
        ws.cell(row=row, column=6).number_format = 'DD.MM.YYYY'
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_trend_analysis_sheet(wb, styles):
    """Create Trend_Analysis sheet."""
    ws = wb["Trend_Analysis"]
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = "Trend Analysis - Quarterly Progress Tracking"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Instructions
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "📊 Track compliance scores quarterly to monitor improvement trends"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
    
    row += 2
    
    # Column headers
    headers = ["Quarter", "Domain 1", "Domain 2", "Domain 3", "Domain 4", "Overall"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Example quarters
    quarters = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025"]
    
    for quarter in quarters:
        ws[f'A{row}'] = quarter
        ws[f'A{row}'].font = Font(bold=True)
        
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["input_cell"])
            cell.number_format = '0.0"%"'
        
        row += 1
    
    # Column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A6'


def create_kpi_dashboard_sheet(wb, styles):
    """Create KPI_Dashboard sheet."""
    ws = wb["KPI_Dashboard"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "KPI Dashboard - Key Performance Indicators"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Column headers
    headers = ["KPI Category", "Metric", "Current Value", "Target", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # KPI categories
    kpis = [
        ("Coverage", "Endpoint DLP coverage %", "", "95%", ""),
        ("Coverage", "Email DLP coverage %", "", "98%", ""),
        ("Coverage", "Cloud DLP coverage %", "", "90%", ""),
        ("", "", "", "", ""),
        ("Effectiveness", "Data loss incidents prevented", "", ">50/month", ""),
        ("Effectiveness", "False positive rate %", "", "<10%", ""),
        ("Effectiveness", "Policy violation detection rate", "", ">95%", ""),
        ("", "", "", "", ""),
        ("Response", "Mean time to detect (MTTD)", "", "<4 hours", ""),
        ("Response", "Mean time to respond (MTTR)", "", "<24 hours", ""),
        ("Response", "Incident resolution rate %", "", ">90%", ""),
        ("", "", "", "", ""),
        ("Compliance", "Policy compliance rate %", "", ">95%", ""),
        ("Compliance", "Audit findings (open)", "", "<5", ""),
        ("Compliance", "Training completion %", "", "100%", ""),
    ]
    
    for category, metric, current, target, status in kpis:
        if category == "":
            row += 1
            continue
        
        ws[f'A{row}'] = category
        ws[f'B{row}'] = metric
        ws[f'C{row}'] = current
        ws[f'D{row}'] = target
        ws[f'E{row}'] = status
        
        if category:
            ws[f'A{row}'].font = Font(bold=True)
        
        apply_style(ws[f'C{row}'], styles["input_cell"])
        apply_style(ws[f'E{row}'], styles["input_cell"])
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_budget_planning_sheet(wb, styles):
    """Create Budget_Planning sheet."""
    ws = wb["Budget_Planning"]
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = "Budget Planning - DLP Program Costs"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Column headers
    headers = ["Cost Category", "Description", "Annual Cost (CHF)", "One-Time Cost (CHF)", "FTE Required", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Budget categories
    budget_items = [
        ("Software Licenses", "DLP solution licenses", "", "", "", ""),
        ("Hardware", "Appliances and servers", "", "", "", ""),
        ("Professional Services", "Implementation and consulting", "", "", "", ""),
        ("Training", "Staff training and awareness", "", "", "", ""),
        ("Personnel", "DLP administrators", "", "", "", ""),
        ("Personnel", "DLP analysts", "", "", "", ""),
        ("Maintenance", "Annual support and updates", "", "", "", ""),
        ("Cloud Services", "Cloud DLP/CASB subscription", "", "", "", ""),
    ]
    
    for category, description, annual, onetime, fte, notes in budget_items:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = description
        ws[f'C{row}'] = annual
        ws[f'D{row}'] = onetime
        ws[f'E{row}'] = fte
        ws[f'F{row}'] = notes
        
        apply_style(ws[f'C{row}'], styles["input_cell"])
        apply_style(ws[f'D{row}'], styles["input_cell"])
        apply_style(ws[f'E{row}'], styles["input_cell"])
        apply_style(ws[f'F{row}'], styles["input_cell"])
        
        row += 1
    
    row += 1
    
    # Totals
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = f"=SUM(C4:C{row-2})"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'] = f"=SUM(D4:D{row-2})"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = '#,##0'
    ws[f'E{row}'] = f"=SUM(E4:E{row-2})"
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'E{row}'].number_format = '0.0'
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_ciso_dpo_approval_sheet(wb, styles):
    """Create CISO_DPO_Approval sheet."""
    ws = wb["CISO_DPO_Approval"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "CISO & DPO Approval Sign-Off"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Approval table
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "APPROVAL SIGNATURES"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    approval_headers = ["Role", "Name", "Signature", "Date", "Comments"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    approvers = [
        "DLP Administrator",
        "Information Security Manager",
        "CISO",
        "DPO",
        "CIO",
        "Executive Management",
    ]
    
    for approver in approvers:
        ws[f'A{row}'] = approver
        for col in ['B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A3'


def create_summary_dashboard_sheet(wb, styles):
    """Create Summary_Dashboard sheet."""
    ws = wb["Summary_Dashboard"]
    
    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "Domain 5 - Master Compliance Dashboard"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Metadata
    row = 3
    ws[f'A{row}'] = "Assessment Date:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 1
    ws[f'A{row}'] = "Completed By:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 1
    ws[f'A{row}'] = "Approved By:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 1
    ws[f'A{row}'] = "Next Review:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = '=B3+90'
    ws[f'B{row}'].number_format = 'DD.MM.YYYY'
    
    row += 2
    
    # Overall compliance section
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL DLP PROGRAM COMPLIANCE"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    # Headers
    headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # Domain scores with external formulas
    domains = [
        ("Domain 1 - Infrastructure", f"='[{WORKBOOK_DOMAIN_1}]Summary_Dashboard'!$B$19", "85%"),
        ("Domain 2 - Classification", f"='[{WORKBOOK_DOMAIN_2}]Summary_Dashboard'!$B$18", "85%"),
        ("Domain 3 - Channels", f"='[{WORKBOOK_DOMAIN_3}]Summary_Dashboard'!$B$23", "85%"),
        ("Domain 4 - Monitoring", f"='[{WORKBOOK_DOMAIN_4}]Summary_Dashboard'!$B$17", "85%"),
    ]
    
    domain_start_row = row
    for domain_name, formula, target in domains:
        ws[f'A{row}'] = domain_name
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0'
        ws[f'C{row}'] = target
        ws[f'D{row}'] = f'=IF(B{row}>=85,"\u2705 Pass",IF(B{row}>=70,"\u26A0\uFE0F Conditional","\u274C Fail"))'
        row += 1
    
    # Overall weighted compliance
    ws[f'A{row}'] = "OVERALL DLP COMPLIANCE"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'B{row}'] = f'=(B{domain_start_row}*0.25+B{domain_start_row+1}*0.20+B{domain_start_row+2}*0.30+B{domain_start_row+3}*0.25)'
    ws[f'B{row}'].font = Font(bold=True, size=14)
    ws[f'B{row}'].number_format = '0.0'
    ws[f'C{row}'] = ">85%"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=IF(B{row}>=85,"\u2705 Pass",IF(B{row}>=70,"\u26A0\uFE0F Conditional","\u274C Fail"))'
    ws[f'D{row}'].font = Font(bold=True, size=14)
    
    row += 2
    
    # Key findings
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "KEY FINDINGS"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    findings = [
        ("Total Gaps Identified:", "=COUNTA(Consolidated_Gap_Analysis!B:B)-1"),
        ("Critical Gaps:", '=COUNTIF(Consolidated_Gap_Analysis!D:D,"Critical")'),
        ("High-Risk Issues:", '=COUNTIF(Risk_Register!H:H,"Critical")+COUNTIF(Risk_Register!H:H,"High")'),
        ("Remediation Progress %:", "=COUNTIF(Remediation_Roadmap!J:J,\"Complete\")/COUNTA(Remediation_Roadmap!J:J)*100"),
        ("Evidence Completeness %:", "=COUNTA(Evidence_Master_Index!B:B)/100*100"),
    ]
    
    for finding_name, formula in findings:
        ws[f'A{row}'] = finding_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        if "%" in finding_name:
            ws[f'B{row}'].number_format = '0.0'
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A3'


# ============================================================================
# SECTION 5: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the workbook."""
    logger.info("=" * 80)
    logger.info(f"Generating {WORKBOOK_ID} - {ASSESSMENT_AREA}")
    logger.info("=" * 80)
    logger.info(f"\n\u2705 Using EXTERNAL FORMULAS with normalized workbook references")
    logger.info(f"\u2705 No manual Find & Replace required!")
    logger.info(f"\nReferenced workbooks:")
    logger.info(f"  - {WORKBOOK_DOMAIN_1}")
    logger.info(f"  - {WORKBOOK_DOMAIN_2}")
    logger.info(f"  - {WORKBOOK_DOMAIN_3}")
    logger.info(f"  - {WORKBOOK_DOMAIN_4}")
    logger.info("")
    
    # Create workbook
    wb = create_workbook()
    
    # Setup styles
    styles = setup_styles()
    
    # Setup data validations
    validations = setup_data_validations(wb)
    
    # Create all sheets
    logger.info("Creating Instructions_Legend...")
    create_instructions_sheet(wb, styles)
    
    logger.info("Creating Executive_Summary...")
    create_executive_summary_sheet(wb, styles)
    
    logger.info("Creating Domain_Rollup_Summary...")
    create_domain_rollup_summary_sheet(wb, styles)
    
    logger.info("Creating Consolidated_Gap_Analysis...")
    create_consolidated_gap_analysis_sheet(wb, styles, validations)
    
    logger.info("Creating Risk_Register...")
    create_risk_register_sheet(wb, styles, validations)
    
    logger.info("Creating Remediation_Roadmap...")
    create_remediation_roadmap_sheet(wb, styles, validations)
    
    logger.info("Creating Evidence_Master_Index...")
    create_evidence_master_index_sheet(wb, styles, validations)
    
    logger.info("Creating Trend_Analysis...")
    create_trend_analysis_sheet(wb, styles)
    
    logger.info("Creating KPI_Dashboard...")
    create_kpi_dashboard_sheet(wb, styles)
    
    logger.info("Creating Budget_Planning...")
    create_budget_planning_sheet(wb, styles)
    
    logger.info("Creating CISO_DPO_Approval...")
    create_ciso_dpo_approval_sheet(wb, styles)
    
    logger.info("Creating Summary_Dashboard...")
    create_summary_dashboard_sheet(wb, styles)
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.12.5_Compliance_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    wb.save(filename)
    
    logger.info("\n" + "=" * 80)
    logger.info(f"\u2705 SUCCESS: {filename}")
    logger.info("=" * 80)
    logger.info(f"\nWorkbook Statistics:")
    logger.info(f"  - Total Sheets: 12")
    logger.info(f"  - Consolidated Gap Analysis: 100 rows")
    logger.info(f"  - Risk Register: 40 rows (5 pre-defined)")
    logger.info(f"  - Remediation Roadmap: 50 rows")
    logger.info(f"  - Evidence Master Index: 100 rows")
    logger.info(f"\n📁 External workbook references:")
    logger.info(f"  - Domain 1: {WORKBOOK_DOMAIN_1}")
    logger.info(f"  - Domain 2: {WORKBOOK_DOMAIN_2}")
    logger.info(f"  - Domain 3: {WORKBOOK_DOMAIN_3}")
    logger.info(f"  - Domain 4: {WORKBOOK_DOMAIN_4}")
    logger.info(f"\n⚙️ NEXT STEPS:")
    logger.info(f"  1. Run normalization script FIRST:")
    logger.info(f"     python3 normalize_assessment_files_a812.py")
    logger.info(f"\n  2. Place this dashboard in Dashboard_Sources folder")
    logger.info(f"     (same folder as normalized assessment workbooks)")
    logger.info(f"\n  3. Open dashboard and click 'Update Links' when prompted")
    logger.info(f"\n  4. Dashboard will auto-populate with compliance data!")
    logger.info("")
    logger.info("=" * 80)
    logger.info(f"\n🎯 Master Compliance Dashboard ready - following A.8.24 pattern!")
    logger.info("=" * 80 + "\n")


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT - COMPLETE COMPLIANCE DASHBOARD GENERATOR
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
