#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.1 - DLP Infrastructure Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 1 of 4: DLP Infrastructure

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific DLP technology stack, deployment architecture,
and infrastructure requirements.

Key customization areas:
1. DLP vendor and product inventory (match your actual deployments)
2. Network architecture integration points (specific to your infrastructure)
3. SIEM/SOC integration requirements (aligned with your security operations)
4. License and capacity thresholds (based on your contracts)
5. Compliance scoring criteria (per your risk framework)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates comprehensive DLP infrastructure assessment workbook for systematic
evaluation of data leakage prevention technology deployment against ISO 27001:2022
Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• DLP technology inventory (network, endpoint, cloud, email, web, database)
• Deployment architecture and coverage analysis
• Technical capabilities assessment (content inspection, pattern matching, ML/AI)
• Vendor management and licensing tracking
• SIEM/SOC integration status
• Gap analysis and remediation planning
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.1_DLP_Infrastructure_YYYYMMDD.xlsx

Sheets (11 total):
1. Instructions_Legend - Assessment guidance and metadata
2. DLP_Technology_Inventory - Complete inventory of all DLP solutions
3. Network_DLP - Network appliances, inline/TAP, protocols covered
4. Endpoint_DLP - Agent deployment (Windows, macOS, Linux, VDI)
5. Email_DLP - Gateway DLP, M365 Purview, Google Workspace DLP
6. Cloud_CASB_DLP - CASB solutions, API-based DLP for SaaS
7. Web_DLP - Proxy-based, SSL inspection, URL filtering integration
8. Database_DAM - Database Activity Monitoring for DLP
9. Gap_Analysis - Infrastructure gaps requiring remediation (40 rows)
10. Evidence_Register - Audit evidence tracking (100 rows)
11. Summary_Dashboard - Compliance metrics and KPIs

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_1_dlp_infrastructure.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.1_DLP_Infrastructure_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organization-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Document all DLP technologies deployed in your environment
    5. Assess deployment coverage and integration status
    6. Collect and document evidence (Evidence_Register sheet)
    7. Complete gap analysis for identified deficiencies
    8. Obtain management approval (Summary_Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 1 of 4 in comprehensive DLP evaluation framework
    Focus: DLP technology infrastructure and deployment architecture
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.2 (Channel Protection Requirements)
    Implementation: ISMS-IMP-A.8.12.1 (DLP Infrastructure Implementation Guide)
    Dashboard:      ISMS-IMP-A.8.12.5 (Compliance Dashboard)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py       ← YOU ARE HERE
       python3 generate_a812_2_data_classification.py
       python3 generate_a812_3_channel_coverage.py
       python3 generate_a812_4_monitoring_response.py
    
    2. Complete assessments (manual - security team, infrastructure team)
    
    3. Normalize filenames for dashboard linking:
       python3 normalize_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Infrastructure Assessment → Normalize → Dashboard → Audit Evidence

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    1 of 4 (DLP Infrastructure)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organization] ISMS Implementation Team
License:              [Organization License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - Swiss FADP (Federal Act on Data Protection)
    - EU GDPR (General Data Protection Regulation)
    - NIST SP 800-53 (Security and Privacy Controls)
    - CIS Controls v8 (Center for Internet Security)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Technology Inventory Prerequisites:
    Before completing this assessment, ensure you have:
    • Current inventory of all DLP technologies deployed
    • Access to DLP vendor documentation and licensing details
    • Network architecture diagrams showing DLP placement
    • Endpoint management system data (agent deployment statistics)
    • SIEM/SOC integration documentation

Assessment Scope:
    This workbook covers ALL DLP infrastructure components:
    • Network-based DLP (inline appliances, TAP/SPAN monitoring)
    • Endpoint DLP agents (Windows, macOS, Linux, VDI)
    • Email gateway DLP (on-premises and cloud)
    • Cloud Access Security Brokers (CASB) with DLP capabilities
    • Web proxy DLP (SSL/TLS inspection, URL filtering)
    • Database Activity Monitoring (DAM) for structured data DLP

Data Classification:
    Generated workbooks contain sensitive organizational security information.
    Handle according to [Organization]'s data classification policy.
    Recommended classification: [Organization] Internal/Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: technology inventory, deployment coverage, integration.

Review Cycle:
    Quarterly: Update technology inventory and deployment coverage
    Annually: Complete full infrastructure reassessment
    Ad-hoc: When new DLP technologies deployed or infrastructure changes

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.1"
WORKBOOK_NAME = "DLP Infrastructure Assessment"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: CONSTANTS & CONFIGURATION
# ============================================================================

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.1"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.2"
ASSESSMENT_AREA = "DLP Infrastructure"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "E7E6E6"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "B4C7E7"         # Light blue (\u1F4CB Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 35


# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    sheets = [
        "Instructions_Legend",
        "DLP_Technology_Inventory",
        "Network_DLP",
        "Endpoint_DLP",
        "Email_DLP",
        "Cloud_CASB_DLP",
        "Web_DLP",
        "Database_DAM",
        "Gap_Analysis",
        "Evidence_Register",
        "Summary_Dashboard",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    return {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
    }


def apply_style(cell, style_dict):
    """
    Apply style template to cell by creating NEW instances.
    NEVER reuse Font/Fill/Border objects across cells.
    """
    if "font" in style_dict:
        cell.font = Font(**{k: v for k, v in style_dict["font"].__dict__.items() if not k.startswith('_')})
    if "fill" in style_dict:
        cell.fill = PatternFill(**{k: v for k, v in style_dict["fill"].__dict__.items() if not k.startswith('_')})
    if "alignment" in style_dict:
        cell.alignment = Alignment(**{k: v for k, v in style_dict["alignment"].__dict__.items() if not k.startswith('_')})
    if "border" in style_dict:
        cell.border = Border(**{k: v for k, v in style_dict["border"].__dict__.items() if not k.startswith('_')})


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """
    Create data validation objects.
    MUST be added to worksheet.add_data_validation() and then cells added to validation.
    """
    return {
        "yes_no_partial": DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            error="Invalid value. Select from dropdown.",
            errorTitle="Invalid Entry"
        ),
        "deployment_type": DataValidation(
            type="list",
            formula1='"Network,Endpoint,Cloud,Email,Web,Database"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_arch": DataValidation(
            type="list",
            formula1='"Inline,Monitor,Hybrid,Cloud-based,TAP,SPAN"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_status": DataValidation(
            type="list",
            formula1='"Production,Staging,Test,Decommissioned"',
            allow_blank=False,
            showDropDown=True
        ),
        "license_type": DataValidation(
            type="list",
            formula1='"Perpetual,Subscription,Open Source,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "support_status": DataValidation(
            type="list",
            formula1='"Active,Expired,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "integration_status": DataValidation(
            type="list",
            formula1='"Integrated,Standalone,Partial"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_mode": DataValidation(
            type="list",
            formula1='"Inline,TAP,SPAN,Cloud Gateway"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_method": DataValidation(
            type="list",
            formula1='"GPO,SCCM,Jamf,Manual,Cloud MDM,Ansible,Puppet"',
            allow_blank=False,
            showDropDown=True
        ),
        "cloud_deployment": DataValidation(
            type="list",
            formula1='"On-Premise,Cloud,Hybrid"',
            allow_blank=False,
            showDropDown=True
        ),
        "casb_integration": DataValidation(
            type="list",
            formula1='"API,Inline,Log Analysis"',
            allow_blank=False,
            showDropDown=True
        ),
        "proxy_type": DataValidation(
            type="list",
            formula1='"Forward,Reverse,Cloud"',
            allow_blank=False,
            showDropDown=True
        ),
        "db_scope": DataValidation(
            type="list",
            formula1='"Production,All,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "risk_level": DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "gap_status": DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Accepted,Closed"',
            allow_blank=False,
            showDropDown=True
        ),
        "evidence_type": DataValidation(
            type="list",
            formula1='"Screenshot,Configuration File,Policy Document,Log Export,Report,Certificate,Email,Meeting Minutes,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "verification_status": DataValidation(
            type="list",
            formula1='"Verified,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
        "approval_status": DataValidation(
            type="list",
            formula1='"Approved,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
    }
    
# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 1
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ISMS Control {WORKBOOK_ID} - {ASSESSMENT_AREA}"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    # Subheader
    ws.merge_cells('A2:H2')
    ws['A2'] = f"ISO/IEC 27001:2022 Control {CONTROL_ID} - Data Leakage Prevention"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Document Info
    info = [
        ("Document ID:", WORKBOOK_ID),
        ("Assessment Area:", ASSESSMENT_AREA),
        ("Related Policy:", RELATED_POLICY),
        ("Version:", WORKBOOK_VERSION),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row = 4
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "USER INPUT" in value:
            ws[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        row += 1
    
    # HOW TO USE THIS WORKBOOK section
    row += 2
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    instructions = [
        "1. Complete each worksheet tab in sequence (Technology Inventory → Network → Endpoint → ... → Summary)",
        "2. Fill ALL yellow-highlighted cells with your organization's information",
        "3. Use dropdown menus where provided (do not type free-form text in dropdown cells)",
        "4. Document all DLP technologies across network, endpoint, cloud, email, web, and database layers",
        "5. Provide evidence IDs for all assessments (format: A812-1-[CATEGORY]-[###])",
        "6. Review Summary Dashboard for overall infrastructure compliance score",
        "7. Identify gaps in Gap_Analysis sheet and create remediation plans",
        "8. Upload evidence files to designated Evidence folder",
        "9. Obtain CISO/DPO approval in Summary Dashboard before finalizing",
        "10. Update this workbook quarterly or upon major infrastructure changes",
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # LEGEND section
    row += 2
    ws[f'A{row}'] = "LEGEND - RESPONSE VALUES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    legend = [
        ("Yes", "Fully implemented and documented", COLOR_COMPLETE),
        ("No", "Not implemented", COLOR_MISSING),
        ("Partial", "Partially implemented (explain in notes)", COLOR_PARTIAL),
        ("Planned", "Scheduled for implementation (provide target date)", COLOR_PLANNED),
        ("N/A", "Not applicable (provide justification)", "FFFFFF"),
    ]
    
    for value, description, color in legend:
        ws[f'A{row}'] = value
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        row += 1
    
    # EVIDENCE NAMING section
    row += 2
    ws[f'A{row}'] = "EVIDENCE NAMING CONVENTION"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    evidence_examples = [
        "Format: A812-1-[CATEGORY]-[###]",
        "Examples:",
        "  \u2022 A812-1-INF-001 = Infrastructure - Technology inventory",
        "  \u2022 A812-1-NET-001 = Network DLP - Configuration screenshot",
        "  \u2022 A812-1-EPT-001 = Endpoint DLP - Deployment report",
        "  \u2022 A812-1-EML-001 = Email DLP - Policy config",
        "  \u2022 A812-1-CLD-001 = Cloud/CASB - CASB dashboard",
    ]
    
    for example in evidence_examples:
        ws[f'A{row}'] = example
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes at row 3
    ws.freeze_panes = 'A3'


def create_technology_inventory(ws, styles):
    """Create DLP Technology Inventory sheet."""
    
    # Header
    ws.merge_cells('A1:P1')
    ws['A1'] = "DLP TECHNOLOGY INVENTORY - All DLP Solutions"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:P2')
    ws['A2'] = "Document ALL DLP technologies deployed (network, endpoint, cloud, email, web, database)"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Technology ID"),
        ("B3", "Technology Name"),
        ("C3", "Deployment Type"),
        ("D3", "Vendor"),
        ("E3", "Version"),
        ("F3", "Deployment Architecture"),
        ("G3", "Deployment Status"),
        ("H3", "License Type"),
        ("I3", "License Expiry"),
        ("J3", "Support Contract"),
        ("K3", "EOL Date"),
        ("L3", "Primary Use Case"),
        ("M3", "Integration Status"),
        ("N3", "SIEM Integration"),
        ("O3", "SOC Integration"),
        ("P3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [15, 25, 20, 20, 15, 20, 18, 18, 15, 15, 15, 30, 18, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples (gray rows)
    examples = [
        ("DLP-001", "Forcepoint DLP", "Network", "Forcepoint", "8.9", "Inline", "Production", "Subscription", "2025-12-31", "Active", "2027-06-30", "Network traffic inspection", "Integrated", "Yes", "Yes", "A812-1-INF-001"),
        ("DLP-002", "Forcepoint Endpoint", "Endpoint", "Forcepoint", "8.9", "Agent-based", "Production", "Subscription", "2025-12-31", "Active", "2027-06-30", "USB, clipboard, print blocking", "Integrated", "Yes", "Yes", "A812-1-INF-002"),
        ("DLP-003", "Microsoft Purview DLP", "Cloud", "Microsoft", "Current", "Cloud-based", "Production", "Subscription", "2026-06-30", "Active", "N/A", "M365 email, OneDrive, Teams", "Integrated", "Partial", "Partial", "A812-1-INF-003"),
        ("DLP-004", "Symantec DLP", "Email", "Broadcom", "15.8", "Inline", "Production", "Perpetual", "N/A", "Active", "2026-12-31", "SMTP gateway inspection", "Integrated", "Yes", "Yes", "A812-1-INF-004"),
        ("DLP-005", "Netskope CASB", "Cloud", "Netskope", "102.1", "Cloud-based", "Production", "Subscription", "2025-09-30", "Active", "N/A", "SaaS application DLP", "Integrated", "Yes", "Yes", "A812-1-INF-005"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows (yellow)
    for r in range(row, row + 25):
        for col_idx in range(1, 17):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Deployment Type (Column C)
    ws.add_data_validation(validations['deployment_type'])
    for r in range(9, 35):
        validations['deployment_type'].add(ws[f'C{r}'])
    
    # Deployment Architecture (Column F)
    ws.add_data_validation(validations['deployment_arch'])
    for r in range(9, 35):
        validations['deployment_arch'].add(ws[f'F{r}'])
    
    # Deployment Status (Column G)
    ws.add_data_validation(validations['deployment_status'])
    for r in range(9, 35):
        validations['deployment_status'].add(ws[f'G{r}'])
    
    # License Type (Column H)
    ws.add_data_validation(validations['license_type'])
    for r in range(9, 35):
        validations['license_type'].add(ws[f'H{r}'])
    
    # Support Contract (Column J)
    ws.add_data_validation(validations['support_status'])
    for r in range(9, 35):
        validations['support_status'].add(ws[f'J{r}'])
    
    # Integration Status (Column M)
    ws.add_data_validation(validations['integration_status'])
    for r in range(9, 35):
        validations['integration_status'].add(ws[f'M{r}'])
    
    # SIEM/SOC Integration (Columns N, O)
    ws.add_data_validation(validations['yes_no_partial'])
    for r in range(9, 35):
        validations['yes_no_partial'].add(ws[f'N{r}'])
        validations['yes_no_partial'].add(ws[f'O{r}'])
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_network_dlp(ws, styles):
    """Create Network DLP Assessment sheet."""
    
    # Header
    ws.merge_cells('A1:N1')
    ws['A1'] = "NETWORK DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:N2')
    ws['A2'] = "Assess network-based DLP appliances (inline, TAP, SPAN configurations)"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Appliance Name"),
        ("B3", "Deployment Mode"),
        ("C3", "Network Segments Covered"),
        ("D3", "Protocols Inspected"),
        ("E3", "SSL/TLS Inspection"),
        ("F3", "Throughput Capacity"),
        ("G3", "Current Utilization %"),
        ("H3", "Content Inspection"),
        ("I3", "Pattern Matching (Regex)"),
        ("J3", "Fingerprinting"),
        ("K3", "Machine Learning/AI"),
        ("L3", "Blocking Capability"),
        ("M3", "High Availability"),
        ("N3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [25, 20, 25, 25, 18, 15, 15, 18, 18, 18, 18, 18, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Forcepoint DLP Gateway", "Inline", "DMZ, Internal", "HTTP, HTTPS, SMTP, FTP", "Yes", "10 Gbps", "45", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "A812-1-NET-001"),
        ("Symantec DLP Network", "TAP", "Internet Gateway", "HTTP, HTTPS, FTP", "Yes", "5 Gbps", "60", "Yes", "Yes", "Yes", "No", "No", "Yes", "A812-1-NET-002"),
        ("McAfee DLP Prevent", "Inline", "Branch Offices", "HTTP, HTTPS", "Partial", "1 Gbps", "30", "Yes", "Yes", "No", "No", "Yes", "No", "A812-1-NET-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 17):
        for col_idx in range(1, 15):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['deployment_mode'])
    for r in range(7, 25):
        validations['deployment_mode'].add(ws[f'B{r}'])
    
    for col in ['E', 'H', 'I', 'J', 'K', 'L']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(7, 25):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    # M column (HA) - Yes/No/N/A only
    ha_val = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False)
    ws.add_data_validation(ha_val)
    for r in range(7, 25):
        ha_val.add(ws[f'M{r}'])
    
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 5: SHEET CREATION FUNCTIONS - PART 2
# ============================================================================

def create_endpoint_dlp(ws, styles):
    """Create Endpoint DLP Assessment sheet."""
    
    # Header
    ws.merge_cells('A1:O1')
    ws['A1'] = "ENDPOINT DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:O2')
    ws['A2'] = "Assess endpoint DLP agent deployment across Windows, macOS, Linux, VDI"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Operating System"),
        ("B3", "OS Version"),
        ("C3", "Agent Name"),
        ("D3", "Agent Version"),
        ("E3", "Deployment Method"),
        ("F3", "Total Endpoints"),
        ("G3", "Agents Deployed"),
        ("H3", "Deployment Coverage %"),
        ("I3", "USB Blocking"),
        ("J3", "Clipboard Monitoring"),
        ("K3", "Print Blocking"),
        ("L3", "Screen Capture Detection"),
        ("M3", "Bluetooth Blocking"),
        ("N3", "Cloud Sync App Monitoring"),
        ("O3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [20, 20, 25, 15, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Windows", "10/11 Enterprise", "Forcepoint Endpoint", "8.9", "GPO", "2000", "1950", "=G4/F4*100", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-EPT-001"),
        ("macOS", "13.x Ventura", "Forcepoint Endpoint", "8.9", "Jamf", "300", "280", "=G5/F5*100", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "A812-1-EPT-002"),
        ("Linux", "Ubuntu 22.04", "Custom Script", "1.0", "Ansible", "50", "45", "=G6/F6*100", "Partial", "No", "Partial", "No", "Partial", "No", "A812-1-EPT-003"),
        ("VDI (Citrix)", "Windows 10", "Forcepoint Endpoint", "8.9", "Image", "500", "500", "=G7/F7*100", "Yes", "Yes", "Yes", "Yes", "N/A", "Yes", "A812-1-EPT-004"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 21):
        for col_idx in range(1, 16):
            cell = ws.cell(row=r, column=col_idx)
            # Column H (Coverage %) gets formula
            if col_idx == 8:
                cell.value = f'=IF(F{r}=0,0,G{r}/F{r}*100)'
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['deployment_method'])
    for r in range(8, 30):
        validations['deployment_method'].add(ws[f'E{r}'])
    
    for col in ['I', 'J', 'K', 'L', 'M', 'N']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(8, 30):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    ws.freeze_panes = 'A4'


def create_email_dlp(ws, styles):
    """Create Email DLP Assessment sheet."""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "EMAIL DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:L2')
    ws['A2'] = "Assess email gateway DLP solutions (Exchange, M365, Gmail)"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Email System"),
        ("B3", "DLP Solution"),
        ("C3", "Deployment Type"),
        ("D3", "SMTP Gateway Protection"),
        ("E3", "Webmail Protection"),
        ("F3", "Attachment Scanning"),
        ("G3", "Encrypted Email Handling"),
        ("H3", "Internal Email Monitoring"),
        ("I3", "External Email Monitoring"),
        ("J3", "Quarantine Capability"),
        ("K3", "Auto-Encryption"),
        ("L3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [25, 25, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Microsoft 365", "Microsoft Purview DLP", "Cloud", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "Yes", "A812-1-EML-001"),
        ("On-Prem Exchange", "Symantec Email DLP", "On-Premise", "Yes", "No", "Yes", "Yes", "Partial", "Yes", "Yes", "Partial", "A812-1-EML-002"),
        ("Gmail Workspace", "Google DLP", "Cloud", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "Partial", "A812-1-EML-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 17):
        for col_idx in range(1, 13):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['cloud_deployment'])
    for r in range(7, 25):
        validations['cloud_deployment'].add(ws[f'C{r}'])
    
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'K']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(7, 25):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    # J column (Quarantine) - Yes/No/Partial/N/A
    quarantine_val = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(quarantine_val)
    for r in range(7, 25):
        quarantine_val.add(ws[f'J{r}'])
    
    ws.freeze_panes = 'A4'


def create_cloud_casb_dlp(ws, styles):
    """Create Cloud/CASB DLP Assessment sheet."""
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "CLOUD / CASB DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Assess Cloud Access Security Broker and cloud-native DLP"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Cloud Service"),
        ("B3", "CASB Solution"),
        ("C3", "Integration Type"),
        ("D3", "Upload Monitoring"),
        ("E3", "Download Monitoring"),
        ("F3", "Sharing Controls"),
        ("G3", "Content Inspection"),
        ("H3", "Real-Time Blocking"),
        ("I3", "Shadow IT Discovery"),
        ("J3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [25, 25, 18, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Microsoft OneDrive", "Netskope", "API", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-CLD-001"),
        ("Dropbox", "Netskope", "API", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-CLD-002"),
        ("Box", "McAfee MVISION", "API", "Yes", "Partial", "Yes", "Yes", "Partial", "Partial", "A812-1-CLD-003"),
        ("Google Drive", "Native Google DLP", "API", "Yes", "Yes", "Yes", "Yes", "Yes", "N/A", "A812-1-CLD-004"),
        ("GitHub", "Netskope", "API", "Yes", "Yes", "No", "Yes", "Yes", "Yes", "A812-1-CLD-005"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 15):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['casb_integration'])
    for r in range(9, 25):
        validations['casb_integration'].add(ws[f'C{r}'])
    
    for col in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(9, 25):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    ws.freeze_panes = 'A4'


def create_web_dlp(ws, styles):
    """Create Web DLP Assessment sheet."""
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "WEB PROXY DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Assess web proxy DLP configurations (forward/reverse proxy, SSL inspection)"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Proxy Solution"),
        ("B3", "Proxy Type"),
        ("C3", "SSL Inspection"),
        ("D3", "HTTP/HTTPS Coverage"),
        ("E3", "Upload Monitoring"),
        ("F3", "Download Monitoring"),
        ("G3", "Cloud Storage Detection"),
        ("H3", "Webmail Monitoring"),
        ("I3", "URL Filtering Integration"),
        ("J3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [25, 18, 18, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Zscaler", "Cloud", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-WEB-001"),
        ("Forcepoint Web", "Forward", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "A812-1-WEB-002"),
        ("Squid Proxy", "Forward", "No", "Partial", "No", "No", "No", "No", "Partial", "A812-1-WEB-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 17):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['proxy_type'])
    for r in range(7, 25):
        validations['proxy_type'].add(ws[f'B{r}'])
    
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(7, 25):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    # I column (URL Filtering) - Yes/No/Partial/N/A
    url_val = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(url_val)
    for r in range(7, 25):
        url_val.add(ws[f'I{r}'])
    
    ws.freeze_panes = 'A4'


def create_database_dam(ws, styles):
    """Create Database Activity Monitoring sheet."""
    
    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "DATABASE ACTIVITY MONITORING (DAM)"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Assess Database Activity Monitoring for DLP use cases"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Database System"),
        ("B3", "DAM Solution"),
        ("C3", "Monitoring Scope"),
        ("D3", "SELECT Query Monitoring"),
        ("E3", "Bulk Export Detection"),
        ("F3", "Privileged User Monitoring"),
        ("G3", "DLP Policy Integration"),
        ("H3", "Alert Triggering"),
        ("I3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [25, 25, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Oracle 19c", "Imperva DAM", "Production", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-DAM-001"),
        ("SQL Server 2019", "IBM Guardium", "Production", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-DAM-002"),
        ("MySQL 8.0", "Native Audit Log", "All", "Partial", "No", "Partial", "No", "Partial", "A812-1-DAM-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 12):
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['db_scope'])
    for r in range(7, 20):
        validations['db_scope'].add(ws[f'C{r}'])
    
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(7, 20):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 6: SHEET CREATION FUNCTIONS - PART 3 (Standard Sheets)
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet (standard across all workbooks)."""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "GAP ANALYSIS - IDENTIFIED DEFICIENCIES"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Document all infrastructure gaps and remediation plans"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Gap ID"),
        ("B3", "Gap Description"),
        ("C3", "Affected Technology"),
        ("D3", "Risk Level"),
        ("E3", "Business Impact"),
        ("F3", "Root Cause"),
        ("G3", "Remediation Plan"),
        ("H3", "Owner"),
        ("I3", "Target Date"),
        ("J3", "Status"),
        ("K3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [18, 35, 25, 15, 30, 30, 35, 20, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated example gaps
    examples = [
        ("GAP-A812-1-001", "Endpoint DLP not deployed to Linux servers", "Endpoint DLP", "High", "Unmonitored data exfiltration from Linux systems", "Budget constraints, vendor compatibility", "Deploy compatible endpoint agent or implement compensating controls", "[Owner]", "[Date]", "Open", "A812-1-GAP-001"),
        ("GAP-A812-1-002", "No SSL inspection on web proxy", "Web DLP", "High", "Encrypted channel bypass for web uploads", "Privacy concerns, certificate management complexity", "Implement SSL inspection with privacy impact assessment", "[Owner]", "[Date]", "Open", "A812-1-GAP-002"),
        ("GAP-A812-1-003", "CASB missing for Dropbox", "Cloud DLP", "Medium", "Shadow IT data leakage via Dropbox", "Unknown usage, not in approved SaaS list", "Deploy CASB coverage or block Dropbox entirely", "[Owner]", "[Date]", "Open", "A812-1-GAP-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 37):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['risk_level'])
    for r in range(7, 45):
        validations['risk_level'].add(ws[f'D{r}'])
    
    ws.add_data_validation(validations['gap_status'])
    for r in range(7, 45):
        validations['gap_status'].add(ws[f'J{r}'])
    
    ws.freeze_panes = 'A4'


def create_evidence_register(ws, styles):
    """Create Evidence Register sheet (standard across all workbooks)."""
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = "EVIDENCE REGISTER - AUDIT TRAIL"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:H2')
    ws['A2'] = "Track all evidence collected during infrastructure assessment"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Evidence ID"),
        ("B3", "Evidence Type"),
        ("C3", "Description"),
        ("D3", "Location/Link"),
        ("E3", "Date Collected"),
        ("F3", "Collected By"),
        ("G3", "Related Requirement"),
        ("H3", "Verification Status"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [18, 20, 35, 30, 15, 20, 25, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # User input rows (100 rows)
    for r in range(4, 104):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['evidence_type'])
    for r in range(4, 104):
        validations['evidence_type'].add(ws[f'B{r}'])
    
    ws.add_data_validation(validations['verification_status'])
    for r in range(4, 104):
        validations['verification_status'].add(ws[f'H{r}'])
    
    ws.freeze_panes = 'A4'


def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard sheet with KPIs and compliance scoring."""
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{ASSESSMENT_AREA} - COMPLIANCE DASHBOARD"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35
    
    # Document info section
    ws['A3'] = "Assessment Date:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = "[USER INPUT]"
    ws['B3'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    ws['A4'] = "Completed By:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "[USER INPUT]"
    ws['B4'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    ws['D3'] = "CISO Approval:"
    ws['D3'].font = Font(bold=True)
    ws['E3'] = "[Dropdown]"
    ws['E3'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    ws['D4'] = "Last Review:"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = "[Date]"
    ws['E4'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    # KPI Section
    ws.merge_cells('A6:F6')
    ws['A6'] = "KEY PERFORMANCE INDICATORS"
    apply_style(ws['A6'], styles["subheader"])
    ws.row_dimensions[6].height = 25
    
    # KPI Headers
    kpi_headers = ['KPI', 'Current Value', 'Target', 'Status', 'Trend', 'Notes']
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # KPIs with formulas
    kpis = [
        ("Total DLP Technologies Deployed", '=COUNTA(DLP_Technology_Inventory!B$4:B$35)-5', "≥5", '=IF(B8>=C8,"\u2705 Pass","\u274C Fail")'),
        ("Production Technologies", '=COUNTIF(DLP_Technology_Inventory!G$4:G$35,"Production")', "≥4", '=IF(B9>=C9,"\u2705 Pass","\u274C Fail")'),
        ("Network DLP Coverage %", '=COUNTIF(Network_DLP!H$4:H$25,"Yes")/20*100', "≥80%", '=IF(B10>=80,"\u2705 Pass","\u274C Fail")'),
        ("Endpoint DLP Coverage %", '=AVERAGE(Endpoint_DLP!H$4:H$30)', "≥95%", '=IF(B11>=95,"\u2705 Pass","\u274C Fail")'),
        ("Email DLP Coverage %", '=COUNTIF(Email_DLP!D$4:D$25,"Yes")/20*100', "≥90%", '=IF(B12>=90,"\u2705 Pass","\u274C Fail")'),
        ("Cloud/CASB DLP Coverage %", '=COUNTIF(Cloud_CASB_DLP!D$4:D$25,"Yes")/20*100', "≥70%", '=IF(B13>=70,"\u2705 Pass","\u274C Fail")'),
        ("Web DLP Coverage %", '=COUNTIF(Web_DLP!C$4:C$25,"Yes")/20*100', "≥75%", '=IF(B14>=75,"\u2705 Pass","\u274C Fail")'),
        ("Database DAM Coverage %", '=COUNTIF(Database_DAM!C$4:C$20,"Production")/15*100', "≥80%", '=IF(B15>=80,"\u2705 Pass","\u274C Fail")'),
        ("SIEM Integration Rate %", '=COUNTIF(DLP_Technology_Inventory!N$4:N$35,"Yes")/B8*100', "≥90%", '=IF(B16>=90,"\u2705 Pass","\u274C Fail")'),
        ("Total Gaps Identified", '=COUNTA(Gap_Analysis!A$4:A$45)-3', "≤10", '=IF(B17<=10,"\u2705 Pass","\u274C Fail")'),
        ("Critical/High Gaps", '=COUNTIFS(Gap_Analysis!D$4:D$45,"Critical")+COUNTIFS(Gap_Analysis!D$4:D$45,"High")', "0", '=IF(B18=0,"\u2705 Pass","\u274C Fail")'),
        ("Overall Infrastructure Compliance %", '=AVERAGE(B10:B16)', "≥85%", '=IF(B19>=85,"\u2705 Compliant",IF(B19>=70,"\u26A0\uFE0F Partial","\u274C Non-Compliant"))'),
    ]
    
    row = 8
    for kpi_name, formula, target, status_formula in kpis:
        ws[f'A{row}'] = kpi_name
        ws[f'B{row}'] = formula
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status_formula
        ws[f'E{row}'] = "→"  # Trend placeholder
        ws[f'F{row}'] = ""   # Notes
        row += 1
    
    # Gap Summary Section
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "GAP SUMMARY BY RISK LEVEL"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    gap_headers = ['Risk Level', 'Count', '% of Total', '', '', '']
    for col_idx, header in enumerate(gap_headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_idx, value=header)
            apply_style(cell, styles["column_header"])
    row += 1
    
    gap_levels = [
        ("Critical", '=COUNTIF(Gap_Analysis!D$4:D$45,"Critical")', '=B{}/B17*100'),
        ("High", '=COUNTIF(Gap_Analysis!D$4:D$45,"High")', '=B{}/B17*100'),
        ("Medium", '=COUNTIF(Gap_Analysis!D$4:D$45,"Medium")', '=B{}/B17*100'),
        ("Low", '=COUNTIF(Gap_Analysis!D$4:D$45,"Low")', '=B{}/B17*100'),
    ]
    
    for level, count_formula, pct_formula in gap_levels:
        ws[f'A{row}'] = level
        ws[f'B{row}'] = count_formula
        ws[f'C{row}'] = pct_formula.format(row)
        row += 1
    
    # Evidence Completeness Section
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "EVIDENCE COMPLETENESS"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    evidence_headers = ['Metric', 'Value', '', '', '', '']
    for col_idx, header in enumerate(evidence_headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_idx, value=header)
            apply_style(cell, styles["column_header"])
    row += 1
    
    evidence_metrics = [
        ("Evidence Items Collected", '=COUNTA(Evidence_Register!A$4:A$105)'),
        ("Evidence Items Verified", '=COUNTIF(Evidence_Register!H$4:H$105,"Verified")'),
        ("Evidence Completeness %", f'=B{row+1}/B{row}*100'),
    ]
    
    for metric, formula in evidence_metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        row += 1
    
    # Approval Section
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "APPROVAL SIGN-OFF"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    approval_headers = ['Approver', 'Name', 'Signature', 'Date', 'Status', '']
    for col_idx, header in enumerate(approval_headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_idx, value=header)
            apply_style(cell, styles["column_header"])
    row += 1
    
    approvers = ["DLP Administrator", "CISO", "DPO"]
    for approver in approvers:
        ws[f'A{row}'] = approver
        ws[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        ws[f'C{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        ws[f'D{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        ws[f'E{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        row += 1
    
    # Add approval status validation
    validations = create_data_validations()
    ws.add_data_validation(validations['approval_status'])
    for r in range(row-3, row):
        validations['approval_status'].add(ws[f'E{r}'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_NARROW
    ws.column_dimensions['F'].width = WIDTH_WIDE
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 7: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    
    logger.info("=" * 78)
    logger.info(f"{WORKBOOK_ID} - {ASSESSMENT_AREA} Generator")
    logger.info(f"ISO/IEC 27001:2022 Control {CONTROL_ID}")
    logger.info("=" * 78)
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("\n[1/11] Creating Instructions & Legend...")
    create_instructions(wb["Instructions_Legend"], styles)
    
    logger.info("[2/11] Creating DLP Technology Inventory...")
    create_technology_inventory(wb["DLP_Technology_Inventory"], styles)
    
    logger.info("[3/11] Creating Network DLP Assessment...")
    create_network_dlp(wb["Network_DLP"], styles)
    
    logger.info("[4/11] Creating Endpoint DLP Assessment...")
    create_endpoint_dlp(wb["Endpoint_DLP"], styles)
    
    logger.info("[5/11] Creating Email DLP Assessment...")
    create_email_dlp(wb["Email_DLP"], styles)
    
    logger.info("[6/11] Creating Cloud/CASB DLP Assessment...")
    create_cloud_casb_dlp(wb["Cloud_CASB_DLP"], styles)
    
    logger.info("[7/11] Creating Web DLP Assessment...")
    create_web_dlp(wb["Web_DLP"], styles)
    
    logger.info("[8/11] Creating Database DAM Assessment...")
    create_database_dam(wb["Database_DAM"], styles)
    
    logger.info("[9/11] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap_Analysis"], styles)
    
    logger.info("[10/11] Creating Evidence Register...")
    create_evidence_register(wb["Evidence_Register"], styles)
    
    logger.info("[11/11] Creating Summary Dashboard...")
    create_summary_dashboard(wb["Summary_Dashboard"], styles)
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.12.1_DLP_Infrastructure_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  \u2022 Instructions & Legend")
    logger.info("  \u2022 DLP Technology Inventory (30 rows: 5 examples + 25 blank)")
    logger.info("  \u2022 Network DLP (20 rows: 3 examples + 17 blank)")
    logger.info("  \u2022 Endpoint DLP (25 rows: 4 examples + 21 blank)")
    logger.info("  \u2022 Email DLP (20 rows: 3 examples + 17 blank)")
    logger.info("  \u2022 Cloud/CASB DLP (20 rows: 5 examples + 15 blank)")
    logger.info("  \u2022 Web DLP (20 rows: 3 examples + 17 blank)")
    logger.info("  \u2022 Database DAM (15 rows: 3 examples + 12 blank)")
    logger.info("  \u2022 Gap Analysis (40 rows: 3 examples + 37 blank)")
    logger.info("  \u2022 Evidence Register (100 rows)")
    logger.info("  \u2022 Summary Dashboard (12 KPIs + compliance tracking)")
    logger.info("\n" + "=" * 78)
    logger.info("NEXT STEPS:")
    logger.info("1. Open the workbook in Excel/LibreOffice")
    logger.info("2. Complete yellow-highlighted cells")
    logger.info("3. Use dropdowns (don't type free-form in dropdown cells)")
    logger.info("4. Gather evidence for all assessments")
    logger.info("5. Review Summary Dashboard for compliance score")
    logger.info("6. Obtain CISO/DPO approval")
    logger.info("=" * 78)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
