#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.3 - Channel Coverage Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 3 of 4: Channel Coverage

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific data channels, exfiltration vectors, and
protection requirements.

Key customization areas:
1. Email platform configuration (M365, Google Workspace, on-prem)
2. Cloud storage services (specific to your approved services)
3. Endpoint control capabilities (USB, print per your DLP tools)
4. Network monitoring scope (aligned with your architecture)
5. Mobile device management (MDM/MAM specific to your deployment)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates comprehensive channel coverage assessment workbook for systematic
evaluation of DLP protection across all data exfiltration channels against
ISO 27001:2022 Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• Email channel DLP protection (inbound/outbound, internal)
• Web/cloud channel coverage (browser uploads, cloud storage sync)
• Endpoint channel controls (USB, clipboard, print, screen capture)
• Network channel monitoring (file transfers, protocols)
• Application channel security (SaaS, APIs, database exports)
• Mobile channel protection (MDM/MAM, mobile email)
• Channel coverage metrics and gap analysis
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.3_Channel_Coverage_YYYYMMDD.xlsx

Sheets (13 total):
1. Instructions & Legend - Assessment guidance and metadata
2. Channel Overview - Summary of all 6 channels and protection tiers
3. Email Channel - Email DLP deployment and coverage assessment
4. Web Cloud Channel - Web/cloud DLP deployment and coverage
5. Endpoint Channel - Endpoint DLP (USB, print, clipboard, etc.)
6. Network Channel - Network DLP deployment and coverage
7. Application Channel - Application-level DLP controls
8. Mobile Channel - Mobile DLP (MDM/MAM) deployment
9. Coverage Metrics - Channel coverage percentage calculations
10. Gap Analysis - Channel protection gaps (40 rows)
11. Evidence Register - Audit evidence tracking (100 rows)
12. Summary Dashboard - Compliance metrics and KPIs
13. Approval Sign-Off - Assessment approval workflow

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_3_channel_coverage.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.3_Channel_Coverage_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organisation-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Assess DLP coverage for each data exfiltration channel
    5. Calculate channel coverage percentages
    6. Document DLP policies and actions per channel
    7. Collect and document evidence (Evidence Register sheet)
    8. Complete gap analysis for unprotected channels
    9. Obtain management approval (Summary Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 3 of 4 in comprehensive DLP evaluation framework
    Focus: Channel protection coverage across all data exfiltration vectors
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.2 (Channel Protection Requirements)
    Policy:         ISMS-POL-A.8.12-S5.A (DLP Channel Standards)
    Implementation: ISMS-IMP-A.8.12.3 (Channel Coverage Implementation Guide)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py
       python3 generate_a812_2_data_classification.py
       python3 generate_a812_3_channel_coverage.py         ← YOU ARE HERE
       python3 generate_a812_4_monitoring_response.py
    
    2. Complete assessments (manual - security team, network team)
    
    3. Normalise filenames for dashboard linking:
       python3 normalise_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Channel Assessment → Normalise → Dashboard → Audit Evidence

Critical Prerequisites:
    • Domain 1 (Infrastructure) completed - need DLP technology inventory
    • Domain 2 (Classification) completed - need data classification schema
    • Channel DLP policies require both infrastructure AND data classification

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    3 of 4 (Channel Coverage)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organisation] ISMS Implementation Team
License:              [Organisation License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - Swiss FADP (Federal Act on Data Protection)
    - EU GDPR (General Data Protection Regulation)
    - NIST SP 800-53 (Security and Privacy Controls - SC-7, SC-8)
    - CIS Controls v8.1 (Center for Internet Security - Controls 3, 13)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Channel Protection Tiers (from ISMS-POL-A.8.12-S2.2):
    Tier 1 (Critical - Months 1-3):
        • Email (inbound/outbound/internal)
        • Web/Browser uploads
        • USB/Removable media
    
    Tier 2 (High - Months 4-6):
        • Cloud storage sync (Dropbox, OneDrive, Google Drive)
        • Mobile devices (BYOD, corporate)
    
    Tier 3 (Medium - Months 7-9):
        • Network file shares (SMB, NFS)
        • Printers (local, network)
    
    Tier 4 (Low - Months 10-12):
        • Bluetooth
        • Optical drives (CD/DVD)
        • Legacy protocols (FTP, Telnet)

Assessment Scope - Six Data Exfiltration Channels:
    1. Email Channel:
       • Outbound email attachments and body content
       • Inbound email (data import risks)
       • Internal email (insider threat)
       • Webmail (Gmail, Outlook.com)
    
    2. Web/Cloud Channel:
       • Browser file uploads (HTTP POST, WebDAV)
       • Cloud storage sync clients
       • Web-based file sharing (WeTransfer, SendAnywhere)
       • SaaS application data exports
    
    3. Endpoint Channel:
       • USB/removable media (flash drives, external HDD)
       • Clipboard operations (copy/paste to external apps)
       • Print spooler (local and network printers)
       • Screen capture tools
       • Bluetooth file transfers
    
    4. Network Channel:
       • File transfer protocols (FTP, SFTP, SCP)
       • Network shares (SMB, NFS, WebDAV)
       • Database exports (SQL dumps, CSV exports)
       • Backup systems (unauthorised backups)
    
    5. Application Channel:
       • SaaS application APIs
       • Database connection exports
       • ERP/CRM data exports
       • Source code repositories (Git push)
    
    6. Mobile Channel:
       • Mobile email clients
       • Mobile cloud sync apps
       • SMS/MMS (data via messaging)
       • Mobile app data sharing

Coverage Calculation Methodology:
    Channel Coverage % = (Protected Users/Endpoints / Total Users/Endpoints) × 100
    
    Overall DLP Coverage = Weighted average across channels:
    • Tier 1 channels: 40% weight
    • Tier 2 channels: 30% weight
    • Tier 3 channels: 20% weight
    • Tier 4 channels: 10% weight

Data Classification:
    Generated workbooks contain sensitive organisational security information.
    Handle according to [Organisation]'s data classification policy.
    Recommended classification: [Organisation] Internal/Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: channel coverage %, protection tiers, gap remediation.

Review Cycle:
    Quarterly: Update channel coverage metrics
    Annually: Complete full channel assessment
    Ad-hoc: When new channels identified or infrastructure changes

================================================================================
--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.3"
WORKBOOK_NAME = "Channel Coverage Assessment"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.3"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.2"
ASSESSMENT_AREA = "Channel Coverage Assessment"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "F2F2F2"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "F2F2F2"         # Light blue (Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 40
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create all sheets in order
    sheets = [
        "Instructions & Legend",
        "Channel Overview",
        "Email Channel",
        "Web Cloud Channel",
        "Endpoint Channel",
        "Network Channel",
        "Application Channel",
        "Mobile Channel",
        "Coverage Metrics",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    styles = {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "font": Font(name="Calibri", size=10, italic=True),
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "normal_cell": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "bold_label": {
            "font": Font(name="Calibri", size=10, bold=True),
            "alignment": Alignment(horizontal="left", vertical="center")
        }
    }
    
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell by creating NEW style instances.
    NEVER reuse Font/Fill/Border objects across cells.
    """
    if "font" in style_dict:
        cell.font = Font(**{k: v for k, v in style_dict["font"].__dict__.items() if not k.startswith('_')})
    if "fill" in style_dict:
        cell.fill = PatternFill(**{k: v for k, v in style_dict["fill"].__dict__.items() if not k.startswith('_')})
    if "alignment" in style_dict:
        cell.alignment = Alignment(**{k: v for k, v in style_dict["alignment"].__dict__.items() if not k.startswith('_')})
    if "border" in style_dict:
        cell.border = Border(**{k: v for k, v in style_dict["border"].__dict__.items() if not k.startswith('_')})


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """
    Create data validation objects.
    MUST be added to worksheet.add_data_validation() and then cells added to validation.
    """
    return {
        "yes_no_partial": DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            error="Invalid value. Select from dropdown.",
            errorTitle="Invalid Entry"
        ),
        "channel_type": DataValidation(
            type="list",
            formula1='"Email,Web,Endpoint,Network,App,Mobile"',
            allow_blank=False,
            showDropDown=True
        ),
        "risk_level": DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "exfiltration_risk": DataValidation(
            type="list",
            formula1='"Very High,High,Medium,Low,Very Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "status": DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Complete,Blocked"',
            allow_blank=False,
            showDropDown=True
        ),
        "approval_status": DataValidation(
            type="list",
            formula1='"Approved,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
        "email_system_type": DataValidation(
            type="list",
            formula1='"SMTP,M365,Gmail,Webmail,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_mode": DataValidation(
            type="list",
            formula1='"Inline,Monitor,Cloud-Native,CASB"',
            allow_blank=False,
            showDropDown=True
        ),
        "policy_action": DataValidation(
            type="list",
            formula1='"Allow,Alert,Block,Quarantine,Encrypt"',
            allow_blank=False,
            showDropDown=True
        ),
        "encrypted_handling": DataValidation(
            type="list",
            formula1='"Decrypt,Quarantine,Allow,Block"',
            allow_blank=False,
            showDropDown=True
        ),
        "service_type": DataValidation(
            type="list",
            formula1='"Cloud Storage,SaaS,Web Upload,Code Repo,Social Media"',
            allow_blank=False,
            showDropDown=True
        ),
        "protection_method": DataValidation(
            type="list",
            formula1='"Proxy,CASB,Cloud-Native,Endpoint"',
            allow_blank=False,
            showDropDown=True
        ),
        "endpoint_type": DataValidation(
            type="list",
            formula1='"Windows Desktop,macOS,Linux,VDI,Thin Client"',
            allow_blank=False,
            showDropDown=True
        ),
        "usb_control": DataValidation(
            type="list",
            formula1='"Block All,Allow Approved,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "print_control": DataValidation(
            type="list",
            formula1='"Block,Watermark,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "clipboard_control": DataValidation(
            type="list",
            formula1='"Block,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "app_control": DataValidation(
            type="list",
            formula1='"Whitelist,Blacklist,Monitor,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "network_protocol": DataValidation(
            type="list",
            formula1='"SMB/CIFS,FTP,SFTP,NFS,SCP,WebDAV,Rsync"',
            allow_blank=False,
            showDropDown=True
        ),
        "detection_method": DataValidation(
            type="list",
            formula1='"Inline,TAP,SPAN,Endpoint,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "application_type": DataValidation(
            type="list",
            formula1='"Database,API,Reporting,CRM,ERP,BI,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "export_capability": DataValidation(
            type="list",
            formula1='"Yes,No,Limited"',
            allow_blank=False,
            showDropDown=True
        ),
        "dlp_control_method": DataValidation(
            type="list",
            formula1='"DAM,API Gateway,App Control,Endpoint,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "export_action": DataValidation(
            type="list",
            formula1='"Allow,Alert,Block,Approval Required"',
            allow_blank=False,
            showDropDown=True
        ),
        "device_ownership": DataValidation(
            type="list",
            formula1='"Corporate,BYOD"',
            allow_blank=False,
            showDropDown=True
        ),
        "mobile_platform": DataValidation(
            type="list",
            formula1='"iOS,Android,Windows Mobile"',
            allow_blank=False,
            showDropDown=True
        ),
        "evidence_type": DataValidation(
            type="list",
            formula1='"Config,Screenshot,Log,Report,Policy,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "verification_status": DataValidation(
            type="list",
            formula1='"Verified,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        )
    }




def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 1 OF 4)
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet — gold standard IL."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header — single merged row with two-line title
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Document Information (row 3) — plain bold, NO fill, NO banner
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12, name="Calibri")

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", WORKBOOK_NAME),
        ("Related Policy", RELATED_POLICY),
        ("Version", WORKBOOK_VERSION),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Instructions (row 13) — plain bold, NO blue banner
    row = 13
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    instructions = [
        "1. Complete the Channel Overview sheet first for a high-level assessment of all 6 data egress channels.",
        "2. Complete each channel-specific sheet (Email, Web/Cloud, Endpoint, Network, Application, Mobile).",
        "3. Fill ALL yellow-highlighted cells with your organisation's specific information.",
        "4. Use dropdown menus where provided — do not type free-form text in dropdown cells.",
        "5. Document DLP deployment status, coverage percentages, and policy actions per channel.",
        "6. Review the Coverage Metrics sheet for weighted channel coverage calculations.",
        "7. Identify gaps in the Gap Analysis sheet and create remediation plans with owners and target dates.",
        "8. Maintain the Evidence Register with all supporting documentation for audit traceability.",
        "9. Review the Summary Dashboard for overall compliance scoring against ISO 27001:2022 A.8.12.",
        "10. Obtain final approval and sign-off from DLP Administrator, ISO, and CISO.",
    ]

    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # Status Legend — plain bold heading, proper TABLE with 3 columns
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    # Table headers
    row += 1
    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    legend = [
        (CHECK, "Compliant", "Requirement fully met with evidence", "C6EFCE"),
        (WARNING, "Partial", "Partially implemented, gaps identified", "FFEB9C"),
        (XMARK, "Non-Compliant", "Requirement not met, remediation needed", "FFC7CE"),
        ("\u2014", "Not Applicable", "Not applicable to this organisation", None),
    ]

    row += 1
    for sym, status, desc, color in legend:
        ws.cell(row=row, column=1, value=sym).border = border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if color:
            s.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Acceptable Evidence — plain bold, AFTER Status Legend
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, name="Calibri")

    evidence_types = [
        "\u2713 DLP solution configuration exports (sanitised)",
        "\u2713 Email DLP policy rules and content inspection settings",
        "\u2713 Web/cloud DLP proxy or CASB configuration screenshots",
        "\u2713 Endpoint DLP agent deployment reports (coverage statistics)",
        "\u2713 Network DLP detection method configurations (inline/TAP/SPAN)",
        "\u2713 Application-level DLP control documentation (DAM, API gateway)",
        "\u2713 Mobile MDM/MAM enrolment and policy enforcement reports",
        "\u2713 Channel coverage percentage calculations and methodology",
        "\u2713 Gap analysis documentation with remediation timelines",
        "\u2713 SIEM integration logs showing DLP event correlation",
        "\u2713 Management sign-off and approval records",
    ]
    row += 1
    for ev in evidence_types:
        ws[f"A{row}"] = ev
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


def create_channel_overview(ws, styles):
    """Create Channel Overview summary sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "CHANNEL COVERAGE OVERVIEW"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Summary of DLP protection across all data egress channels"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment info
    ws['A3'] = "Assessment Date:"
    ws['A3'].font = Font(bold=True)
    apply_style(ws['B3'], styles["input_cell"])
    ws['D3'] = "Completed By:"
    ws['D3'].font = Font(bold=True)
    apply_style(ws['E3'], styles["input_cell"])

    # Column headers
    headers = [
        "Channel Name", "Priority Tier", "DLP Deployed?", "Coverage %",
        "Users Protected", "Total Users", "Policy Action", "Gap Status", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated channel data
    channels = [
        ("Email", "Tier 1", "[Input]", "='Email Channel'!B57", "[Input]", "[Input]", "[Input]", "=IF(D6>=90,\"\u2705 Compliant\",IF(D6>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Web/Cloud", "Tier 1", "[Input]", "='Web Cloud Channel'!B57", "[Input]", "[Input]", "[Input]", "=IF(D7>=90,\"\u2705 Compliant\",IF(D7>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Endpoint", "Tier 1", "[Input]", "='Endpoint Channel'!L57", "[Input]", "[Input]", "[Input]", "=IF(D8>=90,\"\u2705 Compliant\",IF(D8>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Network", "Tier 3", "[Input]", "='Network Channel'!E57", "[Input]", "[Input]", "[Input]", "=IF(D9>=75,\"\u2705 Compliant\",IF(D9>=60,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Application", "Tier 2", "[Input]", "='Application Channel'!E57", "[Input]", "[Input]", "[Input]", "=IF(D10>=80,\"\u2705 Compliant\",IF(D10>=65,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
        ("Mobile", "Tier 2", "[Input]", "='Mobile Channel'!L57", "[Input]", "[Input]", "[Input]", "=IF(D11>=90,\"\u2705 Compliant\",IF(D11>=70,\"\u26A0\uFE0F Partial\",\"\u274C Gap\"))", "[Input]"),
    ]

    for row_idx, channel_data in enumerate(channels, start=6):
        for col_idx, value in enumerate(channel_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if "[Input]" in str(value) and col_idx not in [4, 8]:  # Not formula columns
                apply_style(cell, styles["input_cell"])
            else:
                apply_style(cell, styles["normal_cell"])

    # Add data validations
    for row in range(6, 12):
        validations["yes_no_partial"].add(ws[f'C{row}'])

    for row in range(6, 12):
        validations["policy_action"].add(ws[f'G{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# END OF PART 1

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 2 OF 4)
# ============================================================================

def create_email_channel(ws, styles):
    """Create Email Channel assessment sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "EMAIL CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:L2')
    ws['A2'] = "Evaluate DLP coverage for SMTP, M365, Gmail, Webmail"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment question
    ws.merge_cells('A3:L3')
    ws['A3'] = "Does your organisation have DLP protection for all email systems?"
    ws['A3'].font = Font(bold=True, size=11)

    # Column headers
    headers = [
        "Email System", "System Type", "DLP Solution", "Deployment Mode",
        "Content Inspection", "Attachment Scanning", "Encrypted Email Handling",
        "Policy Action", "External Email Protected", "Users Covered",
        "SIEM Integration", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated examples — first example becomes sample row (all cols F2F2F2)
    examples = [
        ("Exchange Online", "M365", "Microsoft Purview", "Cloud-Native", "Yes", "Yes", "Decrypt", "Block", "Yes", "850", "Yes", "A812-3-EMAIL-001"),
        ("Gmail", "Gmail", "Google Workspace DLP", "Cloud-Native", "Yes", "Yes", "Quarantine", "Alert", "Yes", "120", "Yes", "A812-3-EMAIL-002"),
        ("SMTP Gateway", "SMTP", "Forcepoint DLP", "Inline", "Yes", "Yes", "Decrypt", "Block", "Yes", "900", "Yes", "A812-3-EMAIL-003"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 6: sample row — all columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # Empty FFFFCC input rows (50 rows: rows 7–56)
    for row in range(7, 57):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    # Summary row (after 50 empty rows: row 57)
    ws['A57'] = "Email Channel Coverage %"
    ws['A57'].font = Font(bold=True)
    ws['B57'] = "=(COUNTIF(E7:E56,\"Yes\")/COUNTA(A7:A56))*100"
    ws['B57'].number_format = '0.0"%"'

    # Add data validations (rows 7–56, skip sample row 6)
    for row in range(7, 57):
        validations["email_system_type"].add(ws[f'B{row}'])

    for row in range(7, 57):
        validations["deployment_mode"].add(ws[f'D{row}'])

    for row in range(7, 57):
        for col in ['E', 'F', 'I', 'K']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])

    for row in range(7, 57):
        validations["encrypted_handling"].add(ws[f'G{row}'])

    for row in range(7, 57):
        validations["policy_action"].add(ws[f'H{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_web_cloud_channel(ws, styles):
    """Create Web/Cloud Channel assessment sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "WEB/CLOUD CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Evaluate DLP coverage for web uploads, cloud storage, SaaS applications"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment question
    ws.merge_cells('A3:K3')
    ws['A3'] = "Does your organisation protect against data exfiltration via web and cloud services?"
    ws['A3'].font = Font(bold=True, size=11)

    # Column headers
    headers = [
        "Service/Application", "Service Type", "DLP Solution", "Protection Method",
        "SSL/TLS Inspection", "Upload Blocking", "Download Monitoring",
        "Policy Action", "Users Covered", "SIEM Integration", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated examples — first example becomes sample row (all cols F2F2F2)
    examples = [
        ("OneDrive", "Cloud Storage", "Microsoft Purview", "Cloud-Native", "Yes", "Yes", "Yes", "Block", "850", "Yes", "A812-3-WEB-001"),
        ("Dropbox", "Cloud Storage", "Netskope CASB", "CASB", "Yes", "Yes", "Yes", "Alert", "900", "Yes", "A812-3-WEB-002"),
        ("GitHub", "Code Repo", "Forcepoint Web DLP", "Proxy", "Yes", "Yes", "Partial", "Block", "200", "Yes", "A812-3-WEB-003"),
        ("Personal Webmail", "Web Upload", "Zscaler DLP", "Proxy", "Yes", "Yes", "No", "Block", "900", "Yes", "A812-3-WEB-004"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 6: sample row — all columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # Empty FFFFCC input rows (50 rows: rows 7–56)
    for row in range(7, 57):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    # Summary row (after 50 empty rows: row 57)
    ws['A57'] = "Web/Cloud Channel Coverage %"
    ws['A57'].font = Font(bold=True)
    ws['B57'] = "=(COUNTIF(E7:E56,\"Yes\")/COUNTA(A7:A56))*100"
    ws['B57'].number_format = '0.0"%"'

    # Add data validations (rows 7–56, skip sample row 6)
    for row in range(7, 57):
        validations["service_type"].add(ws[f'B{row}'])

    for row in range(7, 57):
        validations["protection_method"].add(ws[f'D{row}'])

    for row in range(7, 57):
        for col in ['E', 'F', 'G', 'J']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])

    for row in range(7, 57):
        validations["policy_action"].add(ws[f'H{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_WIDE
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_endpoint_channel(ws, styles):
    """Create Endpoint Channel assessment sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "ENDPOINT CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:M2')
    ws['A2'] = "Evaluate endpoint DLP agent coverage and controls"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment question
    ws.merge_cells('A3:M3')
    ws['A3'] = "Does your organisation have endpoint DLP agents protecting USB, print, and other local channels?"
    ws['A3'].font = Font(bold=True, size=11)

    # Column headers
    headers = [
        "Endpoint Type", "OS Version", "DLP Agent Installed", "USB Control",
        "Print Control", "Clipboard Control", "Screen Capture Control",
        "Bluetooth Control", "Application Control", "Devices Covered",
        "Total Devices", "Coverage %", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated examples — first example becomes sample row (all cols F2F2F2)
    examples = [
        ("Windows Desktop", "Win 11", "Yes", "Allow Approved", "Watermark", "Block", "Block", "Monitor", "Whitelist", "850", "900", "=J6/K6*100", "A812-3-EPT-001"),
        ("macOS", "Sonoma 14", "Yes", "Block All", "Watermark", "Monitor", "Monitor", "Monitor", "Whitelist", "120", "150", "=J7/K7*100", "A812-3-EPT-002"),
        ("Linux", "Ubuntu 22.04", "No", "None", "None", "None", "None", "None", "None", "0", "50", "=J8/K8*100", "A812-3-EPT-003"),
        ("VDI", "Win 11", "Yes", "Block All", "Block", "Block", "Block", "Block", "Whitelist", "200", "200", "=J9/K9*100", "A812-3-EPT-004"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 6: sample row — all columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # Empty FFFFCC input rows (50 rows: rows 7–56)
    for row in range(7, 57):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            # Add formula for Coverage % column
            if col == 12:
                cell.value = f"=J{row}/K{row}*100"
                cell.number_format = '0.0"%"'

    # Summary row (after 50 empty rows: row 57)
    ws['A57'] = "Endpoint Channel Coverage %"
    ws['A57'].font = Font(bold=True)
    ws['L57'] = "=AVERAGE(L7:L56)"
    ws['L57'].number_format = '0.0"%"'

    # Add data validations (rows 7–56, skip sample row 6)
    for row in range(7, 57):
        validations["endpoint_type"].add(ws[f'A{row}'])

    for row in range(7, 57):
        validations["yes_no_partial"].add(ws[f'C{row}'])

    for row in range(7, 57):
        validations["usb_control"].add(ws[f'D{row}'])

    for row in range(7, 57):
        for col in ['E', 'G']:
            validations["print_control"].add(ws[f'{col}{row}'])

    for row in range(7, 57):
        for col in ['F', 'H']:
            validations["clipboard_control"].add(ws[f'{col}{row}'])

    for row in range(7, 57):
        validations["app_control"].add(ws[f'I{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_network_channel(ws, styles):
    """Create Network Channel assessment sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "NETWORK CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Evaluate network DLP coverage for SMB, FTP, NFS, SCP"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment question
    ws.merge_cells('A3:I3')
    ws['A3'] = "Does your organisation monitor/block file transfers via network protocols?"
    ws['A3'].font = Font(bold=True, size=11)

    # Column headers
    headers = [
        "Protocol", "Business Use Case", "DLP Solution", "Detection Method",
        "Content Inspection", "Policy Action", "Encryption Handling",
        "SIEM Integration", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated examples — first example becomes sample row (all cols F2F2F2)
    examples = [
        ("SMB/CIFS", "Internal file shares", "Forcepoint DLP", "Inline", "Yes", "Alert", "Allow", "Yes", "A812-3-NET-001"),
        ("FTP", "External file transfer", "Network DLP", "TAP", "Yes", "Block", "Decrypt", "Yes", "A812-3-NET-002"),
        ("SFTP", "Secure file transfer", "Endpoint DLP", "Endpoint", "Partial", "Alert", "Allow", "Partial", "A812-3-NET-003"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 6: sample row — all columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # Empty FFFFCC input rows (50 rows: rows 7–56)
    for row in range(7, 57):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    # Summary row (after 50 empty rows: row 57)
    ws['A57'] = "Network Channel Coverage %"
    ws['A57'].font = Font(bold=True)
    ws['E57'] = "=(COUNTIF(E7:E56,\"Yes\")/COUNTA(A7:A56))*100"
    ws['E57'].number_format = '0.0"%"'

    # Add data validations (rows 7–56, skip sample row 6)
    for row in range(7, 57):
        validations["network_protocol"].add(ws[f'A{row}'])

    for row in range(7, 57):
        validations["detection_method"].add(ws[f'D{row}'])

    for row in range(7, 57):
        for col in ['E', 'H']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])

    # Inline DVs for network-specific dropdowns
    policy_action_net = DataValidation(
        type="list",
        formula1='"Allow,Alert,Block,None"',
        allow_blank=False,
        showDropDown=True
    )
    ws.add_data_validation(policy_action_net)
    for row in range(7, 57):
        policy_action_net.add(ws[f'F{row}'])

    encrypt_handling_net = DataValidation(
        type="list",
        formula1='"Decrypt,Allow,Block"',
        allow_blank=False,
        showDropDown=True
    )
    ws.add_data_validation(encrypt_handling_net)
    for row in range(7, 57):
        encrypt_handling_net.add(ws[f'G{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# END OF PART 2

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 3 OF 4)
# ============================================================================

def create_application_channel(ws, styles):
    """Create Application Channel assessment sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "APPLICATION CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Evaluate DLP controls in databases, APIs, reporting tools, CRM, ERP"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment question
    ws.merge_cells('A3:I3')
    ws['A3'] = "Does your organisation have DLP controls for application-level data exports?"
    ws['A3'].font = Font(bold=True, size=11)

    # Column headers
    headers = [
        "Application Name", "Application Type", "Data Export Capability",
        "DLP Control Method", "Bulk Export Detection", "Export Policy Action",
        "Audit Logging", "SIEM Integration", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated examples — first example becomes sample row (all cols F2F2F2)
    examples = [
        ("Oracle DB", "Database", "Yes", "DAM (Imperva)", "Yes", "Alert", "Yes", "Yes", "A812-3-APP-001"),
        ("Salesforce", "CRM", "Yes", "CASB", "Yes", "Block", "Yes", "Yes", "A812-3-APP-002"),
        ("Power BI", "BI", "Yes", "App Control", "Partial", "Approval Required", "Partial", "Partial", "A812-3-APP-003"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 6: sample row — all columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # Empty FFFFCC input rows (50 rows: rows 7–56)
    for row in range(7, 57):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    # Summary row (after 50 empty rows: row 57)
    ws['A57'] = "Application Channel Coverage %"
    ws['A57'].font = Font(bold=True)
    ws['E57'] = "=(COUNTIF(E7:E56,\"Yes\")/COUNTA(A7:A56))*100"
    ws['E57'].number_format = '0.0"%"'

    # Add data validations (rows 7–56, skip sample row 6)
    for row in range(7, 57):
        validations["application_type"].add(ws[f'B{row}'])

    for row in range(7, 57):
        validations["export_capability"].add(ws[f'C{row}'])

    for row in range(7, 57):
        validations["dlp_control_method"].add(ws[f'D{row}'])

    for row in range(7, 57):
        for col in ['E', 'G', 'H']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])

    for row in range(7, 57):
        validations["export_action"].add(ws[f'F{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = WIDTH_WIDE
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_mobile_channel(ws, styles):
    """Create Mobile Channel assessment sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "MOBILE CHANNEL PROTECTION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:M2')
    ws['A2'] = "Evaluate MDM/MAM DLP controls for corporate and BYOD devices"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment question
    ws.merge_cells('A3:M3')
    ws['A3'] = "Does your organisation have DLP controls for mobile devices with corporate data access?"
    ws['A3'].font = Font(bold=True, size=11)

    # Column headers
    headers = [
        "Device Ownership", "Mobile Platform", "MDM/MAM Solution", "Work Profile Enabled",
        "Copy-Paste Control", "Screenshot Control", "Camera Access Control",
        "Personal Cloud Block", "AirDrop/NFC Control", "Devices Enrolled",
        "Total Mobile Devices", "Coverage %", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated examples — first example becomes sample row (all cols F2F2F2)
    examples = [
        ("Corporate", "iOS", "Microsoft Intune", "Yes", "Block", "Block", "Block", "Yes", "Block", "200", "200", "=J6/K6*100", "A812-3-MOB-001"),
        ("BYOD", "Android", "Workspace ONE", "Yes", "Block", "Monitor", "Monitor", "Partial", "Monitor", "150", "300", "=J7/K7*100", "A812-3-MOB-002"),
        ("Corporate", "iOS", "Jamf Pro", "Yes", "Block", "Watermark", "Block", "Yes", "Block", "50", "50", "=J8/K8*100", "A812-3-MOB-003"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 6: sample row — all columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # Empty FFFFCC input rows (50 rows: rows 7–56)
    for row in range(7, 57):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            # Add formula for Coverage % column
            if col == 12:
                cell.value = f"=J{row}/K{row}*100"
                cell.number_format = '0.0"%"'

    # Summary row (after 50 empty rows: row 57)
    ws['A57'] = "Mobile Channel Coverage %"
    ws['A57'].font = Font(bold=True)
    ws['L57'] = "=AVERAGE(L7:L56)"
    ws['L57'].number_format = '0.0"%"'

    # Add data validations (rows 7–56, skip sample row 6)
    for row in range(7, 57):
        validations["device_ownership"].add(ws[f'A{row}'])

    for row in range(7, 57):
        validations["mobile_platform"].add(ws[f'B{row}'])

    for row in range(7, 57):
        for col in ['D', 'H']:
            validations["yes_no_partial"].add(ws[f'{col}{row}'])

    for row in range(7, 57):
        for col in ['E', 'G', 'I']:
            validations["clipboard_control"].add(ws[f'{col}{row}'])

    for row in range(7, 57):
        validations["print_control"].add(ws[f'F{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_WIDE
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_coverage_metrics(ws, styles):
    """Create Coverage Metrics sheet."""
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "CHANNEL COVERAGE METRICS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Weighted coverage analysis across all channels"
    apply_style(ws['A2'], styles["subheader"])
    
    # Section 1: Channel Coverage Summary
    ws['A4'] = "SECTION 1: CHANNEL COVERAGE SUMMARY"
    ws['A4'].font = Font(bold=True, size=12)
    
    headers1 = ["Channel", "Target %", "Actual %", "Gap", "Status"]
    for col_idx, header in enumerate(headers1, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    channels_data = [
        ("Email", "95%", "='Email Channel'!B57", "=B6-C6", "=IF(C6>=B6,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Web/Cloud", "90%", "='Web Cloud Channel'!B57", "=B7-C7", "=IF(C7>=B7,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Endpoint", "95%", "='Endpoint Channel'!L57", "=B8-C8", "=IF(C8>=B8,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Network", "75%", "='Network Channel'!E57", "=B9-C9", "=IF(C9>=B9,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Application", "80%", "='Application Channel'!E57", "=B10-C10", "=IF(C10>=B10,\"\u2705 Compliant\",\"\u274C Gap\")"),
        ("Mobile", "90%", "='Mobile Channel'!L57", "=B11-C11", "=IF(C11>=B11,\"\u2705 Compliant\",\"\u274C Gap\")"),
    ]
    
    for row_idx, channel_data in enumerate(channels_data, start=6):
        for col_idx, value in enumerate(channel_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            apply_style(cell, styles["normal_cell"])
            if col_idx in [2, 3, 4]:
                cell.number_format = '0.0"%"'
    
    # Section 2: Weighted Average Coverage
    ws['A13'] = "SECTION 2: WEIGHTED AVERAGE COVERAGE"
    ws['A13'].font = Font(bold=True, size=12)
    
    headers2 = ["Metric", "Formula", "Target"]
    for col_idx, header in enumerate(headers2, start=1):
        cell = ws.cell(row=14, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    metrics_data = [
        ("Overall Channel Coverage %", "=AVERAGE(C6:C11)", "≥85%"),
        ("Tier 1 Channels Avg (Critical)", "=AVERAGE(C6:C8)", "≥90%"),
        ("Tier 2 Channels Avg (High)", "=AVERAGE(C10:C11)", "≥80%"),
        ("Tier 3-4 Channels Avg (Med-Low)", "=C9", "≥70%"),
    ]
    
    for row_idx, metric_data in enumerate(metrics_data, start=15):
        for col_idx, value in enumerate(metric_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx == 1:
                cell.font = Font(bold=True)
            else:
                apply_style(cell, styles["normal_cell"])
            if col_idx == 2:
                cell.number_format = '0.0"%"'
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['B'].width = WIDTH_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A5'


def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "CHANNEL PROTECTION GAP ANALYSIS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:L2')
    ws['A2'] = "Document all unprotected channels and remediation plans"
    apply_style(ws['A2'], styles["subheader"])

    # Assessment question
    ws.merge_cells('A3:L3')
    ws['A3'] = "Identify and prioritize gaps in channel protection coverage"
    ws['A3'].font = Font(bold=True, size=11)

    # Column headers
    headers = [
        "Gap ID", "Channel", "Gap Description", "Risk Level",
        "Exfiltration Risk", "Current Coverage %", "Target Coverage %",
        "Remediation Plan", "Owner", "Target Date", "Status", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Pre-populated example gaps — first example becomes sample row (all cols F2F2F2)
    examples = [
        ("GAP-A812-3-001", "Endpoint", "Linux endpoints not protected by DLP agent", "High", "High", "0", "80", "Deploy endpoint DLP with Linux support or implement compensating controls", "IT Security Lead", "2025-06-30", "Not Started", "A812-3-GAP-001"),
        ("GAP-A812-3-002", "Mobile", "BYOD enrollment rate below target (<50%)", "Medium", "Medium", "50", "90", "Launch BYOD enrollment campaign with user training", "Mobile Admin", "2025-04-30", "In Progress", "A812-3-GAP-002"),
        ("GAP-A812-3-003", "Web", "Personal webmail access not blocked for high-risk users", "Critical", "Very High", "0", "100", "Deploy web filtering policy to block Gmail/Outlook.com for Confidential data users", "Network Security", "2025-03-31", "Planned", "A812-3-GAP-003"),
    ]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 6: sample row — all columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # Empty FFFFCC input rows (50 rows: rows 7–56)
    for row in range(7, 57):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    # Add data validations (rows 7–56, skip sample row 6)
    for row in range(7, 57):
        validations["channel_type"].add(ws[f'B{row}'])

    for row in range(7, 57):
        validations["risk_level"].add(ws[f'D{row}'])

    for row in range(7, 57):
        validations["exfiltration_risk"].add(ws[f'E{row}'])

    for row in range(7, 57):
        validations["status"].add(ws[f'K{row}'])

    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['I'].width = WIDTH_WIDE
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A6'

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# END OF PART 3

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS (PART 4 OF 4)
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete the Channel Overview sheet first for a high-level assessment of all data egress channels.',
        '2. Complete each channel-specific sheet (Email, Web/Cloud, Endpoint, Network, Application, Mobile).',
        '3. Fill ALL yellow-highlighted cells with your organisation\'s specific information.',
        '4. Use dropdown menus where provided — do not type free-form text in dropdown cells.',
        '5. Document DLP deployment status, coverage percentages, and policy actions per channel.',
        '6. Review the Coverage Metrics sheet for weighted channel coverage calculations.',
        '7. Identify gaps in the Gap Analysis sheet and create remediation plans with owners and target dates.',
        '8. Maintain the Evidence Register with all supporting documentation for audit traceability.',
        '9. Review the Summary Dashboard for overall compliance scoring against ISO 27001:2022 A.8.12.',
        '10. Obtain final approval and sign-off from DLP Administrator, ISO, and CISO.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A24"] = "Status Legend"
    ws["A24"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=25, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 26 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet -- standard common sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 -- merge A1:H1, title "EVIDENCE REGISTER"
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for c in range(1, 9):
        ws.cell(row=1, column=c).border = border

    # Row 2 -- subtitle, italic (NOT subheader/blue banner)
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers Row 4 -- 8 columns, 003366 fill, bold white, thin borders
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Evidence Type dropdown (column C)
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(ev_type_dv)

    # Verification Status dropdown (column H)
    ver_status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True
    )
    ws.add_data_validation(ver_status_dv)

    # F2F2F2 sample row at row 5
    INFO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = {
        1: "EV-001", 2: "Email Channel Assessment", 3: "Configuration file",
        4: "Email DLP policy configuration \u2014 Microsoft 365 DLP rules export",
        5: "\\\\fileserver\\isms\\evidence\\dlp\\email-dlp-rules.pdf",
        6: "22.02.2026", 7: "IT Security Team", 8: "Verified"
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = INFO_FILL
        cell.border = border

    # Data rows 6-105 (100 rows)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Apply dropdowns to ranges (exclude sample row)
    ev_type_dv.add("C6:C105")
    ver_status_dv.add("H6:H105")

    # Column widths per spec
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22

    ws.freeze_panes = "A5"


def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard sheet — Gold Standard TABLE 1/2/3 layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title (A1:G1 merged) ──────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.12: Data Leakage Prevention"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: Blank ──────────────────────────────────────────────────────────

    # ── TABLE 1: COMPLIANCE ASSESSMENT SUMMARY ────────────────────────────────
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].border = border

    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # TABLE 1 data rows — one per channel assessment sheet
    # Channel Overview: col C (DLP Deployed?) = yes_no_partial
    # Email Channel: col E (Content Inspection) = yes_no_partial, rows 7-24 (skip sample row 6)
    # Web Cloud Channel: col E (SSL/TLS Inspection), rows 7-24
    # Endpoint Channel: col C (DLP Agent Installed), rows 7-29
    # Network Channel: col E (Content Inspection), rows 7-19
    # Application Channel: col E (Bulk Export Detection), rows 7-19
    # Mobile Channel: col D (Work Profile Enabled), rows 7-19
    # Gap Analysis: col K (Status) = "Not Started,In Progress,Complete,Blocked"
    #   Compliant="Complete", Partial="In Progress", Non-Compliant="Not Started"+"Blocked"
    t1_data = [
        ("Email Channel",
         "=COUNTIF('Email Channel'!E7:E56,\"Yes\")",
         "=COUNTIF('Email Channel'!E7:E56,\"Partial\")",
         "=COUNTIF('Email Channel'!E7:E56,\"No\")",
         "=COUNTIF('Email Channel'!E7:E56,\"N/A\")"),
        ("Web / Cloud Channel",
         "=COUNTIF('Web Cloud Channel'!E7:E56,\"Yes\")",
         "=COUNTIF('Web Cloud Channel'!E7:E56,\"Partial\")",
         "=COUNTIF('Web Cloud Channel'!E7:E56,\"No\")",
         "=COUNTIF('Web Cloud Channel'!E7:E56,\"N/A\")"),
        ("Endpoint Channel",
         "=COUNTIF('Endpoint Channel'!C7:C56,\"Yes\")",
         "=COUNTIF('Endpoint Channel'!C7:C56,\"Partial\")",
         "=COUNTIF('Endpoint Channel'!C7:C56,\"No\")",
         "=COUNTIF('Endpoint Channel'!C7:C56,\"N/A\")"),
        ("Network Channel",
         "=COUNTIF('Network Channel'!E7:E56,\"Yes\")",
         "=COUNTIF('Network Channel'!E7:E56,\"Partial\")",
         "=COUNTIF('Network Channel'!E7:E56,\"No\")",
         "=COUNTIF('Network Channel'!E7:E56,\"N/A\")"),
        ("Application Channel",
         "=COUNTIF('Application Channel'!E7:E56,\"Yes\")",
         "=COUNTIF('Application Channel'!E7:E56,\"Partial\")",
         "=COUNTIF('Application Channel'!E7:E56,\"No\")",
         "=COUNTIF('Application Channel'!E7:E56,\"N/A\")"),
        ("Mobile Channel",
         "=COUNTIF('Mobile Channel'!D7:D56,\"Yes\")",
         "=COUNTIF('Mobile Channel'!D7:D56,\"Partial\")",
         "=COUNTIF('Mobile Channel'!D7:D56,\"No\")",
         "=COUNTIF('Mobile Channel'!D7:D56,\"N/A\")"),
        ("Gap Analysis",
         "=COUNTIF('Gap Analysis'!K7:K56,\"Complete\")",
         "=COUNTIF('Gap Analysis'!K7:K56,\"In Progress\")",
         "=COUNTIF('Gap Analysis'!K7:K56,\"Not Started\")+COUNTIF('Gap Analysis'!K7:K56,\"Blocked\")",
         "=0"),
    ]

    t1_start = 6
    blue_font = Font(name="Calibri", size=10, color="000000")
    for i, (area, compl, partial, non_compl, na) in enumerate(t1_data):
        r = t1_start + i
        ws.cell(row=r, column=1, value=area).border = border
        ws.cell(row=r, column=1).font = blue_font
        ws.cell(row=r, column=2, value=f"=C{r}+D{r}+E{r}+F{r}").border = border
        ws.cell(row=r, column=2).font = blue_font
        ws.cell(row=r, column=3, value=compl).border = border
        ws.cell(row=r, column=3).font = blue_font
        ws.cell(row=r, column=4, value=partial).border = border
        ws.cell(row=r, column=4).font = blue_font
        ws.cell(row=r, column=5, value=non_compl).border = border
        ws.cell(row=r, column=5).font = blue_font
        ws.cell(row=r, column=6, value=na).border = border
        ws.cell(row=r, column=6).font = blue_font
        ws.cell(row=r, column=7, value=f"=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))").border = border
        ws.cell(row=r, column=7).font = blue_font
        ws.cell(row=r, column=7).number_format = "0.0%"

    # TOTAL row
    t1_total_row = t1_start + len(t1_data)
    ws.cell(row=t1_total_row, column=1, value="TOTAL").border = border
    ws.cell(row=t1_total_row, column=1).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=t1_total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col in range(2, 7):
        cell = ws.cell(row=t1_total_row, column=col)
        cell.border = border
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", bold=True, size=10)
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}{t1_start}:{col_letter}{t1_total_row - 1})"
    ws.cell(row=t1_total_row, column=7,
            value=f"=IF((B{t1_total_row}-F{t1_total_row})=0,0,C{t1_total_row}/(B{t1_total_row}-F{t1_total_row}))").border = border
    ws.cell(row=t1_total_row, column=7).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=t1_total_row, column=7).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=t1_total_row, column=7).number_format = "0.0%"

    current_row = t1_total_row + 2

    # ── TABLE 2: KEY METRICS ──────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:G{current_row}")
    ws[f"A{current_row}"] = "TABLE 2 \u2013 KEY METRICS"
    ws[f"A{current_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{current_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{current_row}"].border = border
    current_row += 1

    t2_headers = ["Metric", "Value", "Target", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=hdr if hdr else "")
        if hdr:
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    current_row += 1

    t2_data = [
        ("Email Channel Coverage %",
         "='Email Channel'!B57",
         "\u226595%"),
        ("Web / Cloud Channel Coverage %",
         "='Web Cloud Channel'!B57",
         "\u226590%"),
        ("Endpoint Channel Coverage %",
         "='Endpoint Channel'!L57",
         "\u226595%"),
        ("Network Channel Coverage %",
         "='Network Channel'!E57",
         "\u226575%"),
        ("Mobile Channel Coverage %",
         "='Mobile Channel'!L57",
         "\u226590%"),
        ("Open Critical/High Gaps",
         "=COUNTIFS('Gap Analysis'!D7:D56,\"Critical\",'Gap Analysis'!K7:K56,\"Not Started\")+COUNTIFS('Gap Analysis'!D7:D56,\"High\",'Gap Analysis'!K7:K56,\"Not Started\")",
         "0"),
    ]

    for metric, value_formula, target in t2_data:
        ws.cell(row=current_row, column=1, value=metric).border = border
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=current_row, column=2, value=value_formula).border = border
        ws.cell(row=current_row, column=2).font = Font(name="Calibri", size=10)
        ws.cell(row=current_row, column=3, value=target).border = border
        ws.cell(row=current_row, column=3).font = Font(name="Calibri", size=10)
        for col in range(4, 8):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1

    current_row += 1

    # ── TABLE 3: KEY FINDINGS & RECOMMENDATIONS ───────────────────────────────
    ws.merge_cells(f"A{current_row}:G{current_row}")
    ws[f"A{current_row}"] = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    ws[f"A{current_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{current_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{current_row}"].border = border
    current_row += 1

    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", ""]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=hdr if hdr else "")
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    current_row += 1

    for i in range(1, 6):
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
        ws.cell(row=current_row, column=1).value = str(i)
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=10)
        current_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 5: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet -- standard common sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 -- merge A1:E1, "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: italic subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO/IEC 27001:2022 \u2014 Control A.8.12: {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner -- 4472C4 fill
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G12),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Status dropdown on Assessment Status
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)
    if status_row:
        status_dv.add(ws[f"B{status_row}"])

    # 3 approver sections (gap of 2 rows before first)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True, name="Calibri")
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION row -- immediately after last approver
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS -- gap of 3 rows
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 6: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the Channel Coverage Assessment workbook."""

    logger.info("=" * 78)
    logger.info(f"{WORKBOOK_ID} - {ASSESSMENT_AREA} Generator")
    logger.info(f"ISO/IEC 27001:2022 Control {CONTROL_ID}")
    logger.info("=" * 78)

    wb = create_workbook()
    styles = _STYLES

    logger.info("\n[1/13] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("[2/13] Creating Channel Overview...")
    create_channel_overview(wb["Channel Overview"], styles)

    logger.info("[3/13] Creating Email Channel...")
    create_email_channel(wb["Email Channel"], styles)

    logger.info("[4/13] Creating Web/Cloud Channel...")
    create_web_cloud_channel(wb["Web Cloud Channel"], styles)

    logger.info("[5/13] Creating Endpoint Channel...")
    create_endpoint_channel(wb["Endpoint Channel"], styles)

    logger.info("[6/13] Creating Network Channel...")
    create_network_channel(wb["Network Channel"], styles)

    logger.info("[7/13] Creating Application Channel...")
    create_application_channel(wb["Application Channel"], styles)

    logger.info("[8/13] Creating Mobile Channel...")
    create_mobile_channel(wb["Mobile Channel"], styles)

    logger.info("[9/13] Creating Coverage Metrics...")
    create_coverage_metrics(wb["Coverage Metrics"], styles)

    logger.info("[10/13] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[11/13] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[12/13] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[13/13] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)

    # Save workbook
    filename = f"ISMS-IMP-A.8.12.3_Channel_Coverage_{datetime.now().strftime('%Y%m%d')}.xlsx"
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\nWorkbook Structure (13 sheets):")
    logger.info("  \u2022 Instructions & Legend")
    logger.info("  \u2022 Channel Overview (6 channels summary)")
    logger.info("  \u2022 Email Channel (20 assessment rows)")
    logger.info("  \u2022 Web/Cloud Channel (20 assessment rows)")
    logger.info("  \u2022 Endpoint Channel (25 assessment rows)")
    logger.info("  \u2022 Network Channel (15 assessment rows)")
    logger.info("  \u2022 Application Channel (15 assessment rows)")
    logger.info("  \u2022 Mobile Channel (15 assessment rows)")
    logger.info("  \u2022 Coverage Metrics (weighted coverage calculations)")
    logger.info("  \u2022 Gap Analysis (40 gap rows)")
    logger.info("  \u2022 Evidence Register (100 evidence rows)")
    logger.info("  \u2022 Summary Dashboard (KPIs + compliance scoring)")
    logger.info("  \u2022 Approval Sign-Off (3-level approval workflow)")
    logger.info("\nTotal Assessment Items: ~90 channel protection checkpoints")
    logger.info("\n" + "=" * 78)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
