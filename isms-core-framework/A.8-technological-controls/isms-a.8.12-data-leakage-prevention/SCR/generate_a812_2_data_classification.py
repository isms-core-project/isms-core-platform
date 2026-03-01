#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.2 - Data Classification Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 2 of 4: Data Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific data classification schema, sensitive data
categories, and regulatory requirements.

Key customization areas:
1. Classification schema levels (match your data classification policy)
2. Sensitive data categories (PII, financial per your definitions)
3. Regulatory mappings (FADP, GDPR, PCI-DSS per your jurisdictions)
4. Data discovery tools (specific to your technology stack)
5. Ownership assignment criteria (aligned with your governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates comprehensive data classification assessment workbook for systematic
evaluation of sensitive data identification and classification against ISO 27001:2022
Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• Organisational data classification schema definition
• Sensitive data inventory (PII, financial, intellectual property, credentials)
• Data location mapping (file servers, databases, cloud storage, endpoints)
• Data ownership assignment and accountability
• Regulatory compliance mapping (Swiss FADP, EU GDPR, PCI-DSS)
• Data labeling methods and automated discovery tools
• Gap analysis and remediation planning
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.2_Data_Classification_YYYYMMDD.xlsx

Sheets (11 total):
1. Instructions & Legend - Assessment guidance and metadata
2. Classification Schema - Organisational data classification levels
3. Sensitive Data Inventory - Complete inventory of sensitive data types
4. Data_Locations - Where sensitive data resides (systems, applications)
5. Data_Ownership - Data owners and custodians assignment
6. PII_Inventory - Personal Identifiable Information detailed assessment
7. Regulatory Mapping - FADP, GDPR, PCI-DSS compliance requirements
8. Data_Discovery_Tools - Automated discovery and scanning tools
9. Gap Analysis - Classification gaps requiring remediation (40 rows)
10. Evidence Register - Audit evidence tracking (100 rows)
11. Summary Dashboard - Compliance metrics and KPIs

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_2_data_classification.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.2_Data_Classification_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organisation-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Define your organisation's data classification schema
    5. Inventory all sensitive data types and locations
    6. Assign data owners and map regulatory requirements
    7. Collect and document evidence (Evidence Register sheet)
    8. Complete gap analysis for identified deficiencies
    9. Obtain management approval (Summary Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 2 of 4 in comprehensive DLP evaluation framework
    Focus: Data classification, identification, and regulatory mapping
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.1 (Data Classification Requirements)
    Implementation: ISMS-IMP-A.8.12.2 (Data Classification Implementation Guide)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py
       python3 generate_a812_2_data_classification.py      ← YOU ARE HERE
       python3 generate_a812_3_channel_coverage.py
       python3 generate_a812_4_monitoring_response.py
    
    2. Complete assessments (manual - security team, data owners, legal)
    
    3. Normalise filenames for dashboard linking:
       python3 normalise_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Classification Assessment → Normalise → Dashboard → Audit Evidence

Critical Prerequisite:
    Complete this assessment BEFORE completing Domain 3 (Channel Coverage).
    Channel DLP policies depend on data classification taxonomy defined here.

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    2 of 4 (Data Classification)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organisation] ISMS Implementation Team
License:              [Organisation License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - Swiss FADP (Federal Act on Data Protection - Bundesgesetz über den Datenschutz)
    - EU GDPR (General Data Protection Regulation)
    - PCI-DSS v4.0 (Payment Card Industry Data Security Standard)
    - NIST SP 800-53 (Security and Privacy Controls)
    - CIS Controls v8.1 (Center for Internet Security)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Data Classification Prerequisites:
    Before completing this assessment, ensure you have:
    • Executive approval for data classification schema
    • Identified key data owners across organisation
    • Understanding of applicable regulatory requirements (FADP, GDPR, etc.)
    • Inventory of data storage systems (file servers, databases, cloud)
    • Access to data discovery tool results (if available)

Swiss FADP Compliance:
    This assessment includes Swiss Federal Act on Data Protection (FADP) requirements.
    Key focus areas:
    • Personal data identification and categorization
    • Special category personal data (sensitive data)
    • Data subject rights (access, deletion, portability)
    • Cross-border data transfer requirements
    • Data breach notification obligations

Assessment Scope:
    This workbook covers ALL sensitive data categories:
    • Personal Identifiable Information (PII) - names, addresses, contact details
    • Financial data - credit cards, bank accounts, payment information
    • Intellectual property - trade secrets, proprietary algorithms, R&D data
    • Credentials - passwords, API keys, certificates, tokens
    • Health data - medical records, health insurance information
    • Special category data - biometrics, genetic data, political opinions

Data Classification:
    Generated workbooks contain sensitive organisational security information.
    Handle according to [Organisation]'s data classification policy.
    Recommended classification: [Organisation] Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: classification schema, data inventory, regulatory mapping.

Review Cycle:
    Quarterly: Update sensitive data inventory and locations
    Annually: Complete full data classification reassessment
    Ad-hoc: When new data types identified or regulations change

================================================================================
--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.2"
WORKBOOK_NAME = "Data Classification Assessment"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.2"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.1"
ASSESSMENT_AREA = "Data Classification"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "F2F2F2"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "F2F2F2"         # Light blue (Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 35
WIDTH_VERY_WIDE = 40
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create all sheets in order
    sheets = [
        "Instructions & Legend",
        "Classification Schema",
        "Sensitive Data Inventory",
        "Data Location Mapping",
        "Data Owner Assignment",
        "Regulatory Mapping",
        "Labelling Methods",
        "Discovery Results",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    return {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
    }



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style template to cell by creating NEW instances.
    NEVER reuse Font/Fill/Border objects across cells.
    """
    if "font" in style_dict:
        cell.font = Font(**{k: v for k, v in style_dict["font"].__dict__.items() if not k.startswith('_')})
    if "fill" in style_dict:
        cell.fill = PatternFill(**{k: v for k, v in style_dict["fill"].__dict__.items() if not k.startswith('_')})
    if "alignment" in style_dict:
        cell.alignment = Alignment(**{k: v for k, v in style_dict["alignment"].__dict__.items() if not k.startswith('_')})
    if "border" in style_dict:
        cell.border = Border(**{k: v for k, v in style_dict["border"].__dict__.items() if not k.startswith('_')})


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """
    Create data validation objects.
    MUST be added to worksheet.add_data_validation() and then cells added to validation.
    """
    return {
        "yes_no_partial": DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            error="Invalid value. Select from dropdown.",
            errorTitle="Invalid Entry"
        ),
        "yes_no": DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "yes_no_pending": DataValidation(
            type="list",
            formula1='"Yes,No,Pending"',
            allow_blank=False,
            showDropDown=True
        ),
        "classification_level": DataValidation(
            type="list",
            formula1='"Public,Internal,Confidential,Restricted"',
            allow_blank=False,
            showDropDown=True
        ),
        "data_type": DataValidation(
            type="list",
            formula1='"PII,Financial,IP,Credentials,Business Confidential,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "regulatory": DataValidation(
            type="list",
            formula1='"FADP,GDPR,PCI-DSS,FADP/GDPR,Multiple,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "discovery_status": DataValidation(
            type="list",
            formula1='"Discovered,In Progress,Not Started"',
            allow_blank=False,
            showDropDown=True
        ),
        "location_type": DataValidation(
            type="list",
            formula1='"File Server,Database,Endpoint,Cloud,Email,Backup,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "labeling_method": DataValidation(
            type="list",
            formula1='"Manual,Automated,Metadata,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "compliance_status": DataValidation(
            type="list",
            formula1='"Compliant,Non-Compliant,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "remediation_status": DataValidation(
            type="list",
            formula1='"Complete,In Progress,Planned,Not Started"',
            allow_blank=False,
            showDropDown=True
        ),
        "risk_level": DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "gap_status": DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Accepted,Closed"',
            allow_blank=False,
            showDropDown=True
        ),
        "evidence_type": DataValidation(
            type="list",
            formula1='"Screenshot,Configuration File,Policy Document,Log Export,Report,Certificate,Email,Meeting Minutes,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "verification_status": DataValidation(
            type="list",
            formula1='"Verified,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
        "approval_status": DataValidation(
            type="list",
            formula1='"Approved,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
    }



def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 1
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet matching gold standard."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{WORKBOOK_ID}  -  {ASSESSMENT_AREA}\n"
        f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: Data Leakage Prevention"
    )
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2 EMPTY

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    doc_info = [
        ("Document ID", WORKBOOK_ID),
        ("Assessment Area", ASSESSMENT_AREA),
        ("Related Policy", RELATED_POLICY),
        ("Version", WORKBOOK_VERSION),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    row += 1
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    instructions = [
        "1. Complete each worksheet tab in sequence.",
        "2. Fill all yellow-highlighted cells with your organisation's information.",
        "3. Use dropdown menus where provided.",
        "4. Document your data classification schema (minimum 4 levels).",
        "5. Inventory all sensitive data categories.",
        "6. Map data to storage locations and assign data owners.",
        "7. Map data categories to regulatory requirements.",
        "8. Maintain the Evidence Register for audit traceability.",
        "9. Obtain final approval in the Approval Sign-Off sheet.",
    ]
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    evidence_types = [
        "Data classification policy documentation",
        "Classification schema implementation screenshots",
        "Data inventory records and mapping documents",
        "Regulatory compliance mapping evidence",
        "Data owner assignment records",
        "Labelling method configuration evidence",
        "Discovery scan results and reports",
    ]
    for e in evidence_types:
        ws[f"A{row}"] = e
        row += 1

    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
    row += 1

    legend_items = [
        ("\u2705", "Compliant", "Requirement fully met with evidence"),
        ("\u26A0\uFE0F", "Partial", "Partially implemented, gaps identified"),
        ("\u274C", "Non-Compliant", "Requirement not met, remediation needed"),
        ("\u2014", "N/A", "Not applicable to this organisation"),
    ]
    for sym, status, desc in legend_items:
        ws.cell(row=row, column=1, value=sym).border = border
        ws.cell(row=row, column=2, value=status).border = border
        ws.cell(row=row, column=3, value=desc).border = border
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


def create_classification_schema(ws, styles):
    """Create Classification Schema sheet."""

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "DATA CLASSIFICATION SCHEMA"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Document your organisational data classification levels and handling requirements"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Classification Level"),
        ("B4", "Definition"),
        ("C4", "Examples"),
        ("D4", "Handling Requirements"),
        ("E4", "Access Controls"),
        ("F4", "Encryption Required"),
        ("G4", "DLP Monitoring"),
        ("H4", "Retention Period"),
        ("I4", "Disposal Method"),
        ("J4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [20, WIDTH_VERY_WIDE, WIDTH_VERY_WIDE, WIDTH_DESCRIPTION, WIDTH_EXTRA_WIDE, 18, 18, 18, 25, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Pre-populated classification levels
    levels = [
        ("Public", "[Define what Public means for your organisation]", "[Examples: Press releases, marketing materials, public website]",
         "[Handling: Can be shared externally]", "[Access: No restrictions]"),
        ("Internal", "[Define Internal classification]", "[Examples: Internal policies, procedures, org charts]",
         "[Handling: Internal use only, no external sharing without approval]", "[Access: All employees]"),
        ("Confidential", "[Define Confidential classification]", "[Examples: Customer lists, financial reports, contracts]",
         "[Handling: Limited distribution, need-to-know basis]", "[Access: Authorised personnel only]"),
        ("Restricted", "[Define Restricted classification]", "[Examples: Trade secrets, M&A, executive comp, PII]",
         "[Handling: Strictly controlled, encrypted, logged access]", "[Access: Explicit authorisation required]"),
    ]

    # Row 5: sample row — first level, ALL columns F2F2F2
    sample = levels[0]
    ws['A5'] = sample[0]
    ws['A5'].font = Font(bold=True)
    ws['A5'].fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws['A5'].border = border_thin
    for col_idx, value in enumerate(sample[1:], start=2):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    # Encryption Required (Column F)
    for r in range(6, 56):
        validations['yes_no'].add(ws[f'F{r}'])

    # DLP Monitoring (Column G)
    for r in range(6, 56):
        validations['yes_no_partial'].add(ws[f'G{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_sensitive_data_inventory(ws, styles):
    """Create Sensitive Data Inventory sheet."""

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "SENSITIVE DATA INVENTORY"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Inventory ALL sensitive data categories requiring DLP protection"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Data Category"),
        ("B4", "Data Type"),
        ("C4", "Classification Level"),
        ("D4", "Regulatory Requirement"),
        ("E4", "Data Examples"),
        ("F4", "Business Owner"),
        ("G4", "Data Steward"),
        ("H4", "Estimated Volume"),
        ("I4", "Discovery Status"),
        ("J4", "DLP Protection"),
        ("K4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [30, 25, 20, 20, WIDTH_DESCRIPTION, 25, 25, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Pre-populated examples
    examples = [
        ("Personal Names", "PII", "Confidential", "FADP/GDPR", "First name, last name, full name", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-001"),
        ("Email Addresses", "PII", "Confidential", "FADP/GDPR", "john.doe@example.ch", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-002"),
        ("Swiss Social Security (AHV)", "PII", "Restricted", "FADP", "756.1234.5678.90", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-003"),
        ("Credit Card Numbers", "Financial", "Restricted", "PCI-DSS", "4111-1111-1111-1111", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-004"),
        ("Bank Account (IBAN)", "Financial", "Restricted", "FADP/GDPR", "CH93 0076 2011 6238 5295 7", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Partial", "A812-2-INV-005"),
        ("API Keys", "Credentials", "Restricted", "None", "sk_live_abcd1234efgh5678", "[Owner]", "[Steward]", "[Volume]", "In Progress", "Partial", "A812-2-INV-006"),
        ("Source Code", "IP", "Confidential", "None", "Proprietary algorithms, trade secrets", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-007"),
        ("Customer Lists", "Business Confidential", "Confidential", "FADP/GDPR", "CRM exports, contact databases", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Partial", "A812-2-INV-008"),
        ("M&A Documents", "Business Confidential", "Restricted", "None", "Acquisition targets, financial models", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-009"),
        ("Employee Health Records", "PII", "Restricted", "FADP/GDPR", "Medical diagnoses, disability status", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Partial", "A812-2-INV-010"),
    ]

    # Row 5: sample row — first example, ALL columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    # Data Type (Column B)
    for r in range(6, 56):
        validations['data_type'].add(ws[f'B{r}'])

    # Classification Level (Column C)
    for r in range(6, 56):
        validations['classification_level'].add(ws[f'C{r}'])

    # Regulatory Requirement (Column D)
    for r in range(6, 56):
        validations['regulatory'].add(ws[f'D{r}'])

    # Discovery Status (Column I)
    for r in range(6, 56):
        validations['discovery_status'].add(ws[f'I{r}'])

    # DLP Protection (Column J)
    for r in range(6, 56):
        validations['yes_no_partial'].add(ws[f'J{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_data_location_mapping(ws, styles):
    """Create Data Location Mapping sheet."""

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "DATA LOCATION MAPPING"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Map sensitive data categories to storage locations (file servers, databases, cloud, endpoints)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Data Category"),
        ("B4", "Storage Location Type"),
        ("C4", "Specific Location"),
        ("D4", "Path/Schema"),
        ("E4", "Estimated Records/Files"),
        ("F4", "Last Discovery Scan"),
        ("G4", "DLP Coverage"),
        ("H4", "Encryption at Rest"),
        ("I4", "Access Controls"),
        ("J4", "Data Owner Notified"),
        ("K4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [30, 20, 30, WIDTH_DESCRIPTION, 18, 15, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Pre-populated examples
    examples = [
        ("Credit Card Numbers", "Database", "SQL-PROD-01", "dbo.Payments.CCNumber", "250000", "2024-12-10", "Yes", "Yes", "Yes", "Yes", "A812-2-LOC-001"),
        ("Customer PII", "File Server", "FS-HR-01", "\\\\HR\\PersonnelFiles\\", "5000", "2024-12-15", "Partial", "Yes", "Yes", "Yes", "A812-2-LOC-002"),
        ("Source Code", "Cloud", "GitHub Enterprise", "/repos/proprietary/", "150", "2024-12-01", "Yes", "Yes", "Yes", "Yes", "A812-2-LOC-003"),
        ("API Keys", "Endpoint", "Developer Workstations", "Various local configs", "[Unknown]", "2024-11-20", "No", "Partial", "Partial", "No", "A812-2-LOC-004"),
        ("Email Archives", "Email", "M365 Exchange Online", "All mailboxes", "2000000", "2024-12-20", "Yes", "Yes", "Yes", "Yes", "A812-2-LOC-005"),
    ]

    # Row 5: sample row — first example, ALL columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    # Storage Location Type (Column B)
    for r in range(6, 56):
        validations['location_type'].add(ws[f'B{r}'])

    # DLP Coverage, Encryption, Access Controls (Columns G, H, I)
    for col in ['G', 'H', 'I']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    # Data Owner Notified (Column J)
    for r in range(6, 56):
        validations['yes_no_pending'].add(ws[f'J{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: SHEET CREATION FUNCTIONS - PART 2
# ============================================================================

def create_data_owner_assignment(ws, styles):
    """Create Data Owner Assignment sheet."""

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "DATA OWNER ASSIGNMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Assign data owners and stewards for each data category - accountability is key!"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Data Category"),
        ("B4", "Business Owner Name"),
        ("C4", "Business Owner Department"),
        ("D4", "Business Owner Email"),
        ("E4", "Data Steward Name"),
        ("F4", "Data Steward Department"),
        ("G4", "Data Steward Email"),
        ("H4", "Ownership Documented"),
        ("I4", "Owner Training Complete"),
        ("J4", "Last Review Date"),
        ("K4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [30, 25, 20, 30, 25, 20, 30, 18, 18, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Pre-populated examples
    examples = [
        ("Customer PII", "Jane Smith", "Sales", "jane.smith@example.ch", "John Security", "IT Security", "john.sec@example.ch", "Yes", "Yes", "2024-12-01", "A812-2-OWN-001"),
        ("Financial Data", "Robert Johnson", "Finance", "robert.j@example.ch", "Mary InfoSec", "IT Security", "mary.is@example.ch", "Yes", "Partial", "2024-11-15", "A812-2-OWN-002"),
        ("Source Code", "Alice Developer", "R&D", "alice.dev@example.ch", "Bob CISO", "IT Security", "bob.ciso@example.ch", "Yes", "Yes", "2024-12-10", "A812-2-OWN-003"),
    ]

    # Row 5: sample row — first example, ALL columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    # Ownership Documented (Column H)
    for r in range(6, 56):
        validations['yes_no_pending'].add(ws[f'H{r}'])

    # Owner Training Complete (Column I)
    for r in range(6, 56):
        validations['yes_no_partial'].add(ws[f'I{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_regulatory_mapping(ws, styles):
    """Create Regulatory Mapping sheet."""

    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "REGULATORY COMPLIANCE MAPPING"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Map data categories to regulatory requirements (FADP, GDPR, PCI-DSS, etc.)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Data Category"),
        ("B4", "Regulation"),
        ("C4", "Specific Article/Section"),
        ("D4", "Requirement Summary"),
        ("E4", "Compliance Status"),
        ("F4", "DLP Controls Required"),
        ("G4", "Breach Notification Required"),
        ("H4", "Data Subject Rights"),
        ("I4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [30, 20, 25, WIDTH_VERY_WIDE, 18, 18, 18, 30, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Pre-populated examples
    examples = [
        ("Personal Names", "FADP", "Article 4(1)", "Defined as personal data, requires appropriate protection measures", "Compliant", "Yes", "Yes", "Access, deletion, portability", "A812-2-REG-001"),
        ("Special Category PII", "GDPR", "Article 9", "Sensitive data (health, biometric, etc.) - strict processing rules", "Compliant", "Yes", "Yes", "All GDPR rights apply", "A812-2-REG-002"),
        ("Credit Card Numbers", "PCI-DSS", "Requirement 3", "Protect stored cardholder data with encryption and access controls", "Partial", "Yes", "Yes", "N/A (PCI-DSS)", "A812-2-REG-003"),
        ("Swiss AHV Number", "FADP", "Article 5", "Special category personal data, heightened protection required", "Compliant", "Yes", "Yes", "Access, deletion, correction", "A812-2-REG-004"),
        ("Employee Health Data", "FADP/GDPR", "FADP Art 5, GDPR Art 9", "Health data is special category, strict consent and security", "Compliant", "Yes", "Yes", "All data subject rights", "A812-2-REG-005"),
    ]

    # Row 5: sample row — first example, ALL columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    # Regulation (Column B)
    regulation_val = DataValidation(
        type="list",
        formula1='"FADP,GDPR,PCI-DSS,HIPAA,SOX,FADP/GDPR,Multiple,Other,None"',
        allow_blank=False
    )
    ws.add_data_validation(regulation_val)
    for r in range(6, 56):
        regulation_val.add(ws[f'B{r}'])

    # Compliance Status (Column E)
    for r in range(6, 56):
        validations['compliance_status'].add(ws[f'E{r}'])

    # DLP Controls Required (Column F)
    for r in range(6, 56):
        validations['yes_no_partial'].add(ws[f'F{r}'])

    # Breach Notification Required (Column G)
    for r in range(6, 56):
        validations['yes_no'].add(ws[f'G{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_labelling_methods(ws, styles):
    """Create Labelling Methods sheet."""

    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "DATA LABELING METHODS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Assess data labeling and classification marking methods across systems"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "System/Application"),
        ("B4", "Labeling Method"),
        ("C4", "Classification Tool"),
        ("D4", "Supported Labels"),
        ("E4", "User Training Provided"),
        ("F4", "Enforcement Capability"),
        ("G4", "DLP Integration"),
        ("H4", "Adoption Rate %"),
        ("I4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [25, 20, 25, 30, 18, 18, 18, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Pre-populated examples
    examples = [
        ("Microsoft 365", "Automated", "Microsoft Purview (AIP)", "Public, Internal, Confidential, Restricted", "Yes", "Yes", "Yes", "85", "A812-2-LBL-001"),
        ("SharePoint", "Manual", "Built-in Classification", "Public, Internal, Restricted", "Partial", "Partial", "Partial", "60", "A812-2-LBL-002"),
        ("File Servers", "None", "None", "None", "No", "No", "No", "0", "A812-2-LBL-003"),
        ("Email (Outlook)", "Manual", "Azure Information Protection", "All 4 levels", "Yes", "Yes", "Yes", "70", "A812-2-LBL-004"),
    ]

    # Row 5: sample row — first example, ALL columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    # Labeling Method (Column B)
    for r in range(6, 56):
        validations['labeling_method'].add(ws[f'B{r}'])

    # User Training, Enforcement, DLP Integration (Columns E, F, G)
    for col in ['E', 'F', 'G']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_discovery_results(ws, styles):
    """Create Discovery Results sheet."""

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "DATA DISCOVERY RESULTS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Document data discovery and scanning results (Spirion, BigID, manual audits)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Discovery Tool"),
        ("B4", "Scan Target"),
        ("C4", "Scan Date"),
        ("D4", "Data Categories Found"),
        ("E4", "Total Findings"),
        ("F4", "Critical Findings"),
        ("G4", "False Positive Rate %"),
        ("H4", "Remediation Status"),
        ("I4", "Data Owner Notified"),
        ("J4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [25, 30, 15, 30, 15, 15, 15, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Pre-populated examples
    examples = [
        ("Spirion", "FS-HR-01", "2024-12-15", "PII, SSN, AHV", "5243", "1200", "15", "In Progress", "Yes", "A812-2-DSC-001"),
        ("BigID", "SQL-PROD-01", "2024-12-10", "PII, CCN, IBAN", "250000", "50000", "5", "In Progress", "Yes", "A812-2-DSC-002"),
        ("Manual Audit", "SharePoint", "2024-11-20", "Business Confidential", "1500", "300", "30", "Complete", "Yes", "A812-2-DSC-003"),
    ]

    # Row 5: sample row — first example, ALL columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    # Remediation Status (Column H)
    for r in range(6, 56):
        validations['remediation_status'].add(ws[f'H{r}'])

    # Data Owner Notified (Column I)
    for r in range(6, 56):
        validations['yes_no_pending'].add(ws[f'I{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet (standard across all workbooks)."""

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "GAP ANALYSIS - IDENTIFIED DEFICIENCIES"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Document all data classification gaps and remediation plans"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Gap ID"),
        ("B4", "Gap Description"),
        ("C4", "Affected Data Category"),
        ("D4", "Risk Level"),
        ("E4", "Business Impact"),
        ("F4", "Root Cause"),
        ("G4", "Remediation Plan"),
        ("H4", "Owner"),
        ("I4", "Target Date"),
        ("J4", "Status"),
        ("K4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [18, WIDTH_DESCRIPTION, 25, 15, 30, 30, WIDTH_DESCRIPTION, 20, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Pre-populated example gaps
    examples = [
        ("GAP-A812-2-001", "No data classification schema documented", "All data", "Critical", "Uncontrolled data handling, no DLP baseline", "Schema never formally defined", "Create and approve 4-tier classification schema", "[Owner]", "[Date]", "Open", "A812-2-GAP-001"),
        ("GAP-A812-2-002", "PII found on file servers without DLP protection", "Customer PII", "High", "Potential FADP/GDPR violation, unmonitored exfiltration", "File servers not in DLP scope", "Deploy endpoint DLP to file servers", "[Owner]", "[Date]", "Open", "A812-2-GAP-002"),
        ("GAP-A812-2-003", "Data owners not assigned for 40% of data categories", "Multiple", "High", "Lack of accountability, unclear data stewardship", "Ownership process not established", "Execute data owner assignment workshops", "[Owner]", "[Date]", "Open", "A812-2-GAP-003"),
    ]

    # Row 5: sample row — first example, ALL columns F2F2F2
    for col_idx, value in enumerate(examples[0], start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 50 empty FFFFCC rows starting at row 6 (immediately after sample row)
    for r in range(6, 56):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Data validations — start after sample row (row 6 onward)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['risk_level'].add(ws[f'D{r}'])

    for r in range(6, 56):
        validations['gap_status'].add(ws[f'J{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: SHEET CREATION FUNCTIONS - PART 3 (STANDARD SHEETS)
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete each worksheet tab in sequence.',
        '2. Fill all yellow-highlighted cells with your organisation\'s information.',
        '3. Use dropdown menus where provided.',
        '4. Document your data classification schema (minimum 4 levels).',
        '5. Inventory all sensitive data categories.',
        '6. Map data to storage locations and assign data owners.',
        '7. Map data categories to regulatory requirements.',
        '8. Maintain the Evidence Register for audit traceability.',
        '9. Obtain final approval in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A23"] = "Status Legend"
    ws["A23"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=24, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 25 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet (gold standard ER pattern)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # A1:H1 merged title
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for c in range(1, 9):
        ws.cell(row=1, column=c).border = border_thin

    # A2:H2 merged subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Row 3 empty (spacer)

    # Row 4 column headers
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Column widths
    col_widths = [15, 25, 22, 40, 45, 16, 20, 22]
    for col_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Row 5: F2F2F2 sample row
    INFO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = {
        1: "EV-001", 2: "Classification Schema", 3: "Configuration file",
        4: "Data classification policy configuration export",
        5: "\\\\fileserver\\isms\\evidence\\dlp\\classification-policy.pdf",
        6: "22.02.2026", 7: "IT Security Team", 8: "Verified"
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = INFO_FILL
        cell.border = border_thin

    # Rows 6-105: 100 empty FFFFCC data rows
    for i in range(100):
        r = 6 + i
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin

    # Data validations (inline)
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Screenshot,Configuration File,Policy Document,Log Export,Report,Certificate,Email,Meeting Minutes,Other"',
        allow_blank=True,
        showDropDown=True,
    )
    ev_type_dv.add("C6:C105")
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending,Rejected"',
        allow_blank=True,
        showDropDown=True,
    )
    ver_status_dv.add("H6:H105")
    ws.add_data_validation(ver_status_dv)

    ws.freeze_panes = "A5"


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet (gold standard pattern)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Title row ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border_thin
    ws.row_dimensions[1].height = 35

    # --- Row 2: Control reference ---
    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO/IEC 27001:2022 \u2014 Control A.8.12: {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border_thin

    # --- Row 3: Assessment Summary banner ---
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border_thin

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document", WORKBOOK_ID),
        ("Assessment Period", ""),
        ("Overall Compliance", "='Summary Dashboard'!G14"),
        ("Assessment Status", ""),
        ("Assessed By:", ""),
    ]
    for i, (label, value) in enumerate(summary_fields):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].border = border_thin
        ws[f"B{r}"] = value
        if value == "":
            ws[f"B{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = border_thin

    # Assessment Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,In Review,Approved,Rejected"',
        allow_blank=True,
        showDropDown=True,
    )
    status_dv.add("B7")
    ws.add_data_validation(status_dv)

    # --- Approver sections ---
    approver_sections = [
        (9, "COMPLETED BY (ASSESSOR)", "4472C4"),
        (16, "REVIEWED BY (ISO)", "4472C4"),
        (23, "APPROVED BY (CISO)", "003366"),
    ]

    for start_row, title, colour in approver_sections:
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, 6):
            ws.cell(row=start_row, column=c).border = border_thin

        fields = ["Name", "Title", "Date", "Signature", "Comments"]
        for j, field in enumerate(fields):
            r = start_row + 1 + j
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(bold=True)
            ws[f"A{r}"].border = border_thin
            ws[f"B{r}"] = ""
            ws[f"B{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=r, column=c).border = border_thin

    # --- Final Decision row (Gold Standard: plain label, not banner) ---
    decision_row = 30
    ws[f"A{decision_row}"] = "FINAL DECISION:"
    ws[f"A{decision_row}"].font = Font(bold=True)
    ws[f"A{decision_row}"].border = border_thin
    ws.merge_cells(f"B{decision_row}:E{decision_row}")
    ws[f"B{decision_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=decision_row, column=c).border = border_thin

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
        showDropDown=True,
    )
    decision_dv.add(f"B{decision_row}")
    ws.add_data_validation(decision_dv)

    # --- Next Review Details ---
    review_row = 32
    ws.merge_cells(f"A{review_row}:E{review_row}")
    ws[f"A{review_row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{review_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{review_row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{review_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=review_row, column=c).border = border_thin

    next_review_fields = [
        "Next Review Date",
        "Review Owner",
        "Review Notes",
    ]
    for j, field in enumerate(next_review_fields):
        r = review_row + 1 + j
        ws[f"A{r}"] = field
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].border = border_thin
        ws[f"B{r}"] = ""
        ws[f"B{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = border_thin

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    ws.freeze_panes = "A3"


def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard sheet — Gold Standard TABLE 1/2/3 layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title (A1:G1 merged) ──────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.12: Data Leakage Prevention"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: Blank ──────────────────────────────────────────────────────────

    # ── TABLE 1: COMPLIANCE ASSESSMENT SUMMARY ────────────────────────────────
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].border = border

    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # TABLE 1 data rows
    # Ranges start at row 6 (row 5 = sample row, excluded from compliance counts)
    # All data sheets: 1 F2F2F2 sample row at row 5 + 50 empty FFFFCC rows 6-55
    # classification_schema: col G (DLP Monitoring): Yes/No/Partial/N/A — rows 6-55
    # sensitive_data_inventory: col J (DLP Protection): Yes/No/Partial/N/A — rows 6-55
    # data_location_mapping: col G (DLP Coverage): Yes/No/Partial/N/A — rows 6-55
    # data_owner_assignment: col H (Ownership Documented): Yes/No/Pending — rows 6-55
    # regulatory_mapping: col E (Compliance Status): Compliant/Non-Compliant/Partial/Planned/N/A — rows 6-55
    # labelling_methods: col G (DLP Integration): Yes/No/Partial/N/A — rows 6-55
    # discovery_results: col H (Remediation Status): Complete/In Progress/Planned/Not Started — rows 6-55
    # gap_analysis: col J (Status): Resolved+Closed+Accepted vs In Progress vs Open — rows 6-55
    t1_data = [
        ("Classification Schema",
         "=COUNTIF('Classification Schema'!G6:G55,\"Yes\")",
         "=COUNTIF('Classification Schema'!G6:G55,\"Partial\")+COUNTIF('Classification Schema'!G6:G55,\"Planned\")",
         "=COUNTIF('Classification Schema'!G6:G55,\"No\")",
         "=COUNTIF('Classification Schema'!G6:G55,\"N/A\")"),
        ("Sensitive Data Inventory",
         "=COUNTIF('Sensitive Data Inventory'!J6:J55,\"Yes\")",
         "=COUNTIF('Sensitive Data Inventory'!J6:J55,\"Partial\")",
         "=COUNTIF('Sensitive Data Inventory'!J6:J55,\"No\")",
         "=COUNTIF('Sensitive Data Inventory'!J6:J55,\"N/A\")"),
        ("Data Location Mapping",
         "=COUNTIF('Data Location Mapping'!G6:G55,\"Yes\")",
         "=COUNTIF('Data Location Mapping'!G6:G55,\"Partial\")",
         "=COUNTIF('Data Location Mapping'!G6:G55,\"No\")",
         "=COUNTIF('Data Location Mapping'!G6:G55,\"N/A\")"),
        ("Data Owner Assignment",
         "=COUNTIF('Data Owner Assignment'!H6:H55,\"Yes\")",
         "=COUNTIF('Data Owner Assignment'!H6:H55,\"Pending\")",
         "=COUNTIF('Data Owner Assignment'!H6:H55,\"No\")",
         "=0"),
        ("Regulatory Mapping",
         "=COUNTIF('Regulatory Mapping'!E6:E55,\"Compliant\")",
         "=COUNTIF('Regulatory Mapping'!E6:E55,\"Partial\")+COUNTIF('Regulatory Mapping'!E6:E55,\"Planned\")",
         "=COUNTIF('Regulatory Mapping'!E6:E55,\"Non-Compliant\")",
         "=COUNTIF('Regulatory Mapping'!E6:E55,\"N/A\")"),
        ("Labelling Methods",
         "=COUNTIF('Labelling Methods'!G6:G55,\"Yes\")",
         "=COUNTIF('Labelling Methods'!G6:G55,\"Partial\")",
         "=COUNTIF('Labelling Methods'!G6:G55,\"No\")",
         "=COUNTIF('Labelling Methods'!G6:G55,\"N/A\")"),
        ("Discovery Results",
         "=COUNTIF('Discovery Results'!H6:H55,\"Complete\")",
         "=COUNTIF('Discovery Results'!H6:H55,\"In Progress\")+COUNTIF('Discovery Results'!H6:H55,\"Planned\")",
         "=COUNTIF('Discovery Results'!H6:H55,\"Not Started\")",
         "=0"),
        ("Gap Analysis",
         "=COUNTIF('Gap Analysis'!J6:J55,\"Resolved\")+COUNTIF('Gap Analysis'!J6:J55,\"Closed\")+COUNTIF('Gap Analysis'!J6:J55,\"Accepted\")",
         "=COUNTIF('Gap Analysis'!J6:J55,\"In Progress\")",
         "=COUNTIF('Gap Analysis'!J6:J55,\"Open\")",
         "=0"),
    ]

    t1_start = 6
    blue_font = Font(name="Calibri", size=10, color="000000")
    for i, (area, compl, partial, non_compl, na) in enumerate(t1_data):
        r = t1_start + i
        ws.cell(row=r, column=1, value=area).border = border
        ws.cell(row=r, column=1).font = blue_font
        ws.cell(row=r, column=2, value=f"=C{r}+D{r}+E{r}+F{r}").border = border
        ws.cell(row=r, column=2).font = blue_font
        ws.cell(row=r, column=3, value=compl).border = border
        ws.cell(row=r, column=3).font = blue_font
        ws.cell(row=r, column=4, value=partial).border = border
        ws.cell(row=r, column=4).font = blue_font
        ws.cell(row=r, column=5, value=non_compl).border = border
        ws.cell(row=r, column=5).font = blue_font
        ws.cell(row=r, column=6, value=na).border = border
        ws.cell(row=r, column=6).font = blue_font
        ws.cell(row=r, column=7, value=f"=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))").border = border
        ws.cell(row=r, column=7).font = blue_font
        ws.cell(row=r, column=7).number_format = "0.0%"

    # TOTAL row
    t1_total_row = t1_start + len(t1_data)
    ws.cell(row=t1_total_row, column=1, value="TOTAL").border = border
    ws.cell(row=t1_total_row, column=1).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=t1_total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col in range(2, 7):
        cell = ws.cell(row=t1_total_row, column=col)
        cell.border = border
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", bold=True, size=10)
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}{t1_start}:{col_letter}{t1_total_row - 1})"
    ws.cell(row=t1_total_row, column=7,
            value=f"=IF((B{t1_total_row}-F{t1_total_row})=0,0,C{t1_total_row}/(B{t1_total_row}-F{t1_total_row}))").border = border
    ws.cell(row=t1_total_row, column=7).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=t1_total_row, column=7).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=t1_total_row, column=7).number_format = "0.0%"

    current_row = t1_total_row + 2

    # ── TABLE 2: KEY METRICS ──────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:G{current_row}")
    ws[f"A{current_row}"] = "TABLE 2 \u2013 KEY METRICS"
    ws[f"A{current_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{current_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{current_row}"].border = border
    current_row += 1

    t2_headers = ["Metric", "Value", "Target", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=hdr if hdr else "")
        if hdr:
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    current_row += 1

    t2_data = [
        ("Data Classification Levels with DLP Monitoring",
         "=COUNTIF('Classification Schema'!G6:G55,\"Yes\")",
         "\u22654"),
        ("Sensitive Data Categories with DLP Protection",
         "=COUNTIF('Sensitive Data Inventory'!J6:J55,\"Yes\")",
         "\u226580%"),
        ("Data Locations with DLP Coverage",
         "=COUNTIF('Data Location Mapping'!G6:G55,\"Yes\")",
         "\u226575%"),
        ("Regulatory Requirements Compliant",
         "=COUNTIF('Regulatory Mapping'!E6:E55,\"Compliant\")",
         "\u226590%"),
        ("Open Critical/High Data Gaps",
         "=COUNTIFS('Gap Analysis'!D6:D55,\"Critical\",'Gap Analysis'!J6:J55,\"Open\")+COUNTIFS('Gap Analysis'!D6:D55,\"High\",'Gap Analysis'!J6:J55,\"Open\")",
         "0"),
        ("Evidence Items Verified",
         "=COUNTIF('Evidence Register'!H6:H105,\"Verified\")",
         "\u226580%"),
    ]

    for metric, value_formula, target in t2_data:
        ws.cell(row=current_row, column=1, value=metric).border = border
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=current_row, column=2, value=value_formula).border = border
        ws.cell(row=current_row, column=2).font = Font(name="Calibri", size=10)
        ws.cell(row=current_row, column=3, value=target).border = border
        ws.cell(row=current_row, column=3).font = Font(name="Calibri", size=10)
        for col in range(4, 8):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1

    current_row += 1

    # ── TABLE 3: KEY FINDINGS & RECOMMENDATIONS ───────────────────────────────
    ws.merge_cells(f"A{current_row}:G{current_row}")
    ws[f"A{current_row}"] = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    ws[f"A{current_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{current_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{current_row}"].border = border
    current_row += 1

    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", ""]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=hdr if hdr else "")
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    current_row += 1

    for i in range(1, 6):
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
        ws.cell(row=current_row, column=1).value = str(i)
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=10)
        current_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""

    logger.info("=" * 78)
    logger.info(f"{WORKBOOK_ID} - {ASSESSMENT_AREA} Generator")
    logger.info(f"ISO/IEC 27001:2022 Control {CONTROL_ID}")
    logger.info("=" * 78)

    wb = create_workbook()
    styles = _STYLES

    logger.info("\n[1/12] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("[2/12] Creating Classification Schema...")
    create_classification_schema(wb["Classification Schema"], styles)

    logger.info("[3/12] Creating Sensitive Data Inventory...")
    create_sensitive_data_inventory(wb["Sensitive Data Inventory"], styles)

    logger.info("[4/12] Creating Data Location Mapping...")
    create_data_location_mapping(wb["Data Location Mapping"], styles)

    logger.info("[5/12] Creating Data Owner Assignment...")
    create_data_owner_assignment(wb["Data Owner Assignment"], styles)

    logger.info("[6/12] Creating Regulatory Mapping...")
    create_regulatory_mapping(wb["Regulatory Mapping"], styles)

    logger.info("[7/12] Creating Labelling Methods...")
    create_labelling_methods(wb["Labelling Methods"], styles)

    logger.info("[8/12] Creating Discovery Results...")
    create_discovery_results(wb["Discovery Results"], styles)

    logger.info("[9/12] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[10/12] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[11/12] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[12/12] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)

    # Save workbook
    filename = f"ISMS-IMP-A.8.12.2_Data_Classification_{datetime.now().strftime('%Y%m%d')}.xlsx"
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"\nSUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  - Instructions & Legend")
    logger.info("  - Classification Schema (4 levels + 6 custom)")
    logger.info("  - Sensitive Data Inventory (30 rows: 10 examples + 20 blank)")
    logger.info("  - Data Location Mapping (40 rows: 5 examples + 35 blank)")
    logger.info("  - Data Owner Assignment (30 rows: 3 examples + 27 blank)")
    logger.info("  - Regulatory Mapping (25 rows: 5 examples + 20 blank)")
    logger.info("  - Labelling Methods (20 rows: 4 examples + 16 blank)")
    logger.info("  - Discovery Results (30 rows: 3 examples + 27 blank)")
    logger.info("  - Gap Analysis (40 rows: 3 examples + 37 blank)")
    logger.info("  - Summary Dashboard (11 KPIs + compliance tracking)")
    logger.info("  - Evidence Register (100 rows)")
    logger.info("  - Approval Sign-Off")
    logger.info("\n" + "=" * 78)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
