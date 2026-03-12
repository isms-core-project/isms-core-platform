#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.1 - DLP Infrastructure Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 1 of 4: DLP Infrastructure

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific DLP technology stack, deployment architecture,
and infrastructure requirements.

Key customization areas:
1. DLP vendor and product inventory (match your actual deployments)
2. Network architecture integration points (specific to your infrastructure)
3. SIEM/SOC integration requirements (aligned with your security operations)
4. License and capacity thresholds (based on your contracts)
5. Compliance scoring criteria (per your risk framework)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates comprehensive DLP infrastructure assessment workbook for systematic
evaluation of data leakage prevention technology deployment against ISO 27001:2022
Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• DLP technology inventory (network, endpoint, cloud, email, web, database)
• Deployment architecture and coverage analysis
• Technical capabilities assessment (content inspection, pattern matching, ML/AI)
• Vendor management and licensing tracking
• SIEM/SOC integration status
• Gap analysis and remediation planning
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.1_DLP_Infrastructure_YYYYMMDD.xlsx

Sheets (11 total):
1. Instructions & Legend - Assessment guidance and metadata
2. DLP Technology Inventory - Complete inventory of all DLP solutions
3. Network DLP - Network appliances, inline/TAP, protocols covered
4. Endpoint DLP - Agent deployment (Windows, macOS, Linux, VDI)
5. Email DLP - Gateway DLP, M365 Purview, Google Workspace DLP
6. Cloud CASB DLP - CASB solutions, API-based DLP for SaaS
7. Web DLP - Proxy-based, SSL inspection, URL filtering integration
8. Database DAM - Database Activity Monitoring for DLP
9. Gap Analysis - Infrastructure gaps requiring remediation (40 rows)
10. Evidence Register - Audit evidence tracking (100 rows)
11. Summary Dashboard - Compliance metrics and KPIs

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_1_dlp_infrastructure.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.1_DLP_Infrastructure_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organisation-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Document all DLP technologies deployed in your environment
    5. Assess deployment coverage and integration status
    6. Collect and document evidence (Evidence Register sheet)
    7. Complete gap analysis for identified deficiencies
    8. Obtain management approval (Summary Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 1 of 4 in comprehensive DLP evaluation framework
    Focus: DLP technology infrastructure and deployment architecture
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.2 (Channel Protection Requirements)
    Implementation: ISMS-IMP-A.8.12.1 (DLP Infrastructure Implementation Guide)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py       ← YOU ARE HERE
       python3 generate_a812_2_data_classification.py
       python3 generate_a812_3_channel_coverage.py
       python3 generate_a812_4_monitoring_response.py
    
    2. Complete assessments (manual - security team, infrastructure team)
    
    3. Normalise filenames for dashboard linking:
       python3 normalise_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Infrastructure Assessment → Normalise → Dashboard → Audit Evidence

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    1 of 4 (DLP Infrastructure)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organisation] ISMS Implementation Team
License:              [Organisation License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - Swiss FADP (Federal Act on Data Protection)
    - EU GDPR (General Data Protection Regulation)
    - NIST SP 800-53 (Security and Privacy Controls)
    - CIS Controls v8.1 (Center for Internet Security)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Technology Inventory Prerequisites:
    Before completing this assessment, ensure you have:
    • Current inventory of all DLP technologies deployed
    • Access to DLP vendor documentation and licensing details
    • Network architecture diagrams showing DLP placement
    • Endpoint management system data (agent deployment statistics)
    • SIEM/SOC integration documentation

Assessment Scope:
    This workbook covers ALL DLP infrastructure components:
    • Network-based DLP (inline appliances, TAP/SPAN monitoring)
    • Endpoint DLP agents (Windows, macOS, Linux, VDI)
    • Email gateway DLP (on-premises and cloud)
    • Cloud Access Security Brokers (CASB) with DLP capabilities
    • Web proxy DLP (SSL/TLS inspection, URL filtering)
    • Database Activity Monitoring (DAM) for structured data DLP

Data Classification:
    Generated workbooks contain sensitive organisational security information.
    Handle according to [Organisation]'s data classification policy.
    Recommended classification: [Organisation] Internal/Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: technology inventory, deployment coverage, integration.

Review Cycle:
    Quarterly: Update technology inventory and deployment coverage
    Annually: Complete full infrastructure reassessment
    Ad-hoc: When new DLP technologies deployed or infrastructure changes

================================================================================
--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.1"
WORKBOOK_NAME = "DLP Infrastructure Assessment"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.1"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.2"
ASSESSMENT_AREA = "DLP Infrastructure"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "F2F2F2"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "F2F2F2"         # Light blue (Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 35
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create all sheets in order
    sheets = [
        "Instructions & Legend",
        "DLP Technology Inventory",
        "Network DLP",
        "Endpoint DLP",
        "Email DLP",
        "Cloud CASB DLP",
        "Web DLP",
        "Database DAM",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    return {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
    }



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style template to cell by creating NEW instances.
    NEVER reuse Font/Fill/Border objects across cells.
    """
    if "font" in style_dict:
        cell.font = Font(**{k: v for k, v in style_dict["font"].__dict__.items() if not k.startswith('_')})
    if "fill" in style_dict:
        cell.fill = PatternFill(**{k: v for k, v in style_dict["fill"].__dict__.items() if not k.startswith('_')})
    if "alignment" in style_dict:
        cell.alignment = Alignment(**{k: v for k, v in style_dict["alignment"].__dict__.items() if not k.startswith('_')})
    if "border" in style_dict:
        cell.border = Border(**{k: v for k, v in style_dict["border"].__dict__.items() if not k.startswith('_')})


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """
    Create data validation objects.
    MUST be added to worksheet.add_data_validation() and then cells added to validation.
    """
    return {
        "yes_no_partial": DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            error="Invalid value. Select from dropdown.",
            errorTitle="Invalid Entry"
        ),
        "deployment_type": DataValidation(
            type="list",
            formula1='"Network,Endpoint,Cloud,Email,Web,Database"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_arch": DataValidation(
            type="list",
            formula1='"Inline,Monitor,Hybrid,Cloud-based,TAP,SPAN"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_status": DataValidation(
            type="list",
            formula1='"Production,Staging,Test,Decommissioned"',
            allow_blank=False,
            showDropDown=True
        ),
        "license_type": DataValidation(
            type="list",
            formula1='"Perpetual,Subscription,Open Source,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "support_status": DataValidation(
            type="list",
            formula1='"Active,Expired,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "integration_status": DataValidation(
            type="list",
            formula1='"Integrated,Standalone,Partial"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_mode": DataValidation(
            type="list",
            formula1='"Inline,TAP,SPAN,Cloud Gateway"',
            allow_blank=False,
            showDropDown=True
        ),
        "deployment_method": DataValidation(
            type="list",
            formula1='"GPO,SCCM,Jamf,Manual,Cloud MDM,Ansible,Puppet"',
            allow_blank=False,
            showDropDown=True
        ),
        "cloud_deployment": DataValidation(
            type="list",
            formula1='"On-Premise,Cloud,Hybrid"',
            allow_blank=False,
            showDropDown=True
        ),
        "casb_integration": DataValidation(
            type="list",
            formula1='"API,Inline,Log Analysis"',
            allow_blank=False,
            showDropDown=True
        ),
        "proxy_type": DataValidation(
            type="list",
            formula1='"Forward,Reverse,Cloud"',
            allow_blank=False,
            showDropDown=True
        ),
        "db_scope": DataValidation(
            type="list",
            formula1='"Production,All,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "risk_level": DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "gap_status": DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Accepted,Closed"',
            allow_blank=False,
            showDropDown=True
        ),
        "evidence_type": DataValidation(
            type="list",
            formula1='"Screenshot,Configuration File,Policy Document,Log Export,Report,Certificate,Email,Meeting Minutes,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "verification_status": DataValidation(
            type="list",
            formula1='"Verified,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
        "approval_status": DataValidation(
            type="list",
            formula1='"Approved,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
    }
    


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 1
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet matching gold standard."""
    # Header - two-line with \n, A1:G1 merge
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{WORKBOOK_ID}  -  {ASSESSMENT_AREA}\n"
        f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: Data Leakage Prevention"
    )
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2 must be EMPTY (no subheader)

    # Document Information at A3
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    doc_info = [
        ("Document ID", WORKBOOK_ID),
        ("Assessment Area", ASSESSMENT_AREA),
        ("Related Policy", RELATED_POLICY),
        ("Version", WORKBOOK_VERSION),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    row += 1  # gap
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    instructions = [
        "1. Complete each worksheet tab in sequence.",
        "2. Fill all yellow-highlighted cells with your organisation's information.",
        "3. Use dropdown menus where provided.",
        "4. Document all DLP technologies across network, endpoint, cloud, email, web, and database layers.",
        "5. Provide evidence references for all assessments.",
        "6. Review Summary Dashboard for overall compliance score.",
        "7. Identify gaps in Gap Analysis sheet and create remediation plans.",
        "8. Maintain the Evidence Register for audit traceability.",
        "9. Obtain final approval in the Approval Sign-Off sheet.",
    ]
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1  # gap
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    evidence_types = [
        "Configuration files or screenshots",
        "Network scan results",
        "System documentation and architecture diagrams",
        "Vendor specifications and encryption statements",
        "Certificate inventory and key management records",
        "Audit logs and SIEM evidence",
        "Compliance reports and exception approvals",
    ]
    for e in evidence_types:
        ws[f"A{row}"] = e
        row += 1

    row += 1  # gap
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    # Table headers with D9D9D9 fill
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
    row += 1

    legend_items = [
        ("\u2705", "Compliant", "Requirement fully met with evidence"),
        ("\u26A0\uFE0F", "Partial", "Partially implemented, gaps identified"),
        ("\u274C", "Non-Compliant", "Requirement not met, remediation needed"),
        ("\u2014", "N/A", "Not applicable to this organisation"),
    ]
    for sym, status, desc in legend_items:
        ws.cell(row=row, column=1, value=sym).border = border
        ws.cell(row=row, column=2, value=status).border = border
        ws.cell(row=row, column=3, value=desc).border = border
        row += 1

    # Column widths and freeze
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


def create_technology_inventory(ws, styles):
    """Create DLP Technology Inventory sheet."""

    # Header
    ws.merge_cells('A1:P1')
    ws['A1'] = "DLP TECHNOLOGY INVENTORY - ALL DLP SOLUTIONS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:P2')
    ws['A2'] = "Document ALL DLP technologies deployed (network, endpoint, cloud, email, web, database)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Technology ID"),
        ("B4", "Technology Name"),
        ("C4", "Deployment Type"),
        ("D4", "Vendor"),
        ("E4", "Version"),
        ("F4", "Deployment Architecture"),
        ("G4", "Deployment Status"),
        ("H4", "License Type"),
        ("I4", "License Expiry"),
        ("J4", "Support Contract"),
        ("K4", "EOL Date"),
        ("L4", "Primary Use Case"),
        ("M4", "Integration Status"),
        ("N4", "SIEM Integration"),
        ("O4", "SOC Integration"),
        ("P4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [15, 25, 20, 20, 15, 20, 18, 18, 15, 15, 15, 30, 18, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example data
    examples = [
        ("DLP-001", "Forcepoint DLP", "Network", "Forcepoint", "8.9", "Inline", "Production", "Subscription", "2025-12-31", "Active", "2027-06-30", "Network traffic inspection", "Integrated", "Yes", "Yes", "A812-1-INF-001"),
        ("DLP-002", "Forcepoint Endpoint", "Endpoint", "Forcepoint", "8.9", "Agent-based", "Production", "Subscription", "2025-12-31", "Active", "2027-06-30", "USB, clipboard, print blocking", "Integrated", "Yes", "Yes", "A812-1-INF-002"),
        ("DLP-003", "Microsoft Purview DLP", "Cloud", "Microsoft", "Current", "Cloud-based", "Production", "Subscription", "2026-06-30", "Active", "N/A", "M365 email, OneDrive, Teams", "Integrated", "Partial", "Partial", "A812-1-INF-003"),
        ("DLP-004", "Symantec DLP", "Email", "Broadcom", "15.8", "Inline", "Production", "Perpetual", "N/A", "Active", "2026-12-31", "SMTP gateway inspection", "Integrated", "Yes", "Yes", "A812-1-INF-004"),
        ("DLP-005", "Netskope CASB", "Cloud", "Netskope", "102.1", "Cloud-based", "Production", "Subscription", "2025-09-30", "Active", "N/A", "SaaS application DLP", "Integrated", "Yes", "Yes", "A812-1-INF-005"),
    ]

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey, shows "how to fill in")
    sample = examples[0]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 17):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Add data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    # Deployment Type (Column C)
    for r in range(6, 56):
        validations['deployment_type'].add(ws[f'C{r}'])

    # Deployment Architecture (Column F)
    for r in range(6, 56):
        validations['deployment_arch'].add(ws[f'F{r}'])

    # Deployment Status (Column G)
    for r in range(6, 56):
        validations['deployment_status'].add(ws[f'G{r}'])

    # License Type (Column H)
    for r in range(6, 56):
        validations['license_type'].add(ws[f'H{r}'])

    # Support Contract (Column J)
    for r in range(6, 56):
        validations['support_status'].add(ws[f'J{r}'])

    # Integration Status (Column M)
    for r in range(6, 56):
        validations['integration_status'].add(ws[f'M{r}'])

    # SIEM/SOC Integration (Columns N, O)
    for r in range(6, 56):
        validations['yes_no_partial'].add(ws[f'N{r}'])
        validations['yes_no_partial'].add(ws[f'O{r}'])

    # Freeze panes at A5 (below headers row 4)
    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_network_dlp(ws, styles):
    """Create Network DLP Assessment sheet."""

    # Header
    ws.merge_cells('A1:N1')
    ws['A1'] = "NETWORK DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:N2')
    ws['A2'] = "Assess network-based DLP appliances (inline, TAP, SPAN configurations)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Appliance Name"),
        ("B4", "Deployment Mode"),
        ("C4", "Network Segments Covered"),
        ("D4", "Protocols Inspected"),
        ("E4", "SSL/TLS Inspection"),
        ("F4", "Throughput Capacity"),
        ("G4", "Current Utilization %"),
        ("H4", "Content Inspection"),
        ("I4", "Pattern Matching (Regex)"),
        ("J4", "Fingerprinting"),
        ("K4", "Machine Learning/AI"),
        ("L4", "Blocking Capability"),
        ("M4", "High Availability"),
        ("N4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [25, 20, 25, 25, 18, 15, 15, 18, 18, 18, 18, 18, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example data
    examples = [
        ("Forcepoint DLP Gateway", "Inline", "DMZ, Internal", "HTTP, HTTPS, SMTP, FTP", "Yes", "10 Gbps", "45", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "A812-1-NET-001"),
        ("Symantec DLP Network", "TAP", "Internet Gateway", "HTTP, HTTPS, FTP", "Yes", "5 Gbps", "60", "Yes", "Yes", "Yes", "No", "No", "Yes", "A812-1-NET-002"),
        ("McAfee DLP Prevent", "Inline", "Branch Offices", "HTTP, HTTPS", "Partial", "1 Gbps", "30", "Yes", "Yes", "No", "No", "Yes", "No", "A812-1-NET-003"),
    ]

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey)
    sample = examples[0]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 15):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['deployment_mode'].add(ws[f'B{r}'])

    for col in ['E', 'H', 'I', 'J', 'K', 'L']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    # M column (HA) - Yes/No/N/A only
    ha_val = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False)
    ws.add_data_validation(ha_val)
    for r in range(6, 56):
        ha_val.add(ws[f'M{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: SHEET CREATION FUNCTIONS - PART 2
# ============================================================================

def create_endpoint_dlp(ws, styles):
    """Create Endpoint DLP Assessment sheet."""

    # Header
    ws.merge_cells('A1:O1')
    ws['A1'] = "ENDPOINT DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:O2')
    ws['A2'] = "Assess endpoint DLP agent deployment across Windows, macOS, Linux, VDI"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Operating System"),
        ("B4", "OS Version"),
        ("C4", "Agent Name"),
        ("D4", "Agent Version"),
        ("E4", "Deployment Method"),
        ("F4", "Total Endpoints"),
        ("G4", "Agents Deployed"),
        ("H4", "Deployment Coverage %"),
        ("I4", "USB Blocking"),
        ("J4", "Clipboard Monitoring"),
        ("K4", "Print Blocking"),
        ("L4", "Screen Capture Detection"),
        ("M4", "Bluetooth Blocking"),
        ("N4", "Cloud Sync App Monitoring"),
        ("O4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [20, 20, 25, 15, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example data (Coverage % formula excluded — col index 7, 0-based)
    examples_data = [
        ("Windows", "10/11 Enterprise", "Forcepoint Endpoint", "8.9", "GPO", "2000", "1950", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-EPT-001"),
        ("macOS", "13.x Ventura", "Forcepoint Endpoint", "8.9", "Jamf", "300", "280", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "A812-1-EPT-002"),
        ("Linux", "Ubuntu 22.04", "Custom Script", "1.0", "Ansible", "50", "45", "Partial", "No", "Partial", "No", "Partial", "No", "A812-1-EPT-003"),
        ("VDI", "Windows 10", "Forcepoint Endpoint", "8.9", "Image", "500", "500", "Yes", "Yes", "Yes", "Yes", "N/A", "Yes", "A812-1-EPT-004"),
    ]
    # examples_data cols: 0=OS, 1=OSVer, 2=Agent, 3=AgentVer, 4=Method, 5=Total, 6=Deployed,
    #                     7=USB, 8=Clipboard, 9=Print, 10=Screen, 11=BT, 12=Cloud, 13=EvidID
    # Columns in sheet: A=1..G=7 are data cols 0-6, H=8 is Coverage % formula, I-O = data cols 7-13

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey)
    # Col H (Coverage %) gets the formula referencing row 5
    sample = examples_data[0]
    col_map = list(range(1, 8)) + [None] + list(range(9, 16))  # None = col H (formula)
    for col_idx in range(1, 16):
        cell = ws.cell(row=5, column=col_idx)
        if col_idx == 8:
            cell.value = '=IF(F5=0,0,G5/F5*100)'
        elif col_idx <= 7:
            cell.value = sample[col_idx - 1]
        else:
            cell.value = sample[col_idx - 2]  # skip H (Coverage %) in data array
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 16):
            cell = ws.cell(row=r, column=col_idx)
            # Column H (Coverage %) gets formula in empty rows
            if col_idx == 8:
                cell.value = f'=IF(F{r}=0,0,G{r}/F{r}*100)'
            apply_style(cell, styles["input_cell"])

    # Data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['deployment_method'].add(ws[f'E{r}'])

    for col in ['I', 'J', 'K', 'L', 'M', 'N']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_email_dlp(ws, styles):
    """Create Email DLP Assessment sheet."""

    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "EMAIL DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:L2')
    ws['A2'] = "Assess email gateway DLP solutions (Exchange, M365, Gmail)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Email System"),
        ("B4", "DLP Solution"),
        ("C4", "Deployment Type"),
        ("D4", "SMTP Gateway Protection"),
        ("E4", "Webmail Protection"),
        ("F4", "Attachment Scanning"),
        ("G4", "Encrypted Email Handling"),
        ("H4", "Internal Email Monitoring"),
        ("I4", "External Email Monitoring"),
        ("J4", "Quarantine Capability"),
        ("K4", "Auto-Encryption"),
        ("L4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [25, 25, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example data
    examples = [
        ("Microsoft 365", "Microsoft Purview DLP", "Cloud", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "Yes", "A812-1-EML-001"),
        ("On-Prem Exchange", "Symantec Email DLP", "On-Premise", "Yes", "No", "Yes", "Yes", "Partial", "Yes", "Yes", "Partial", "A812-1-EML-002"),
        ("Gmail Workspace", "Google DLP", "Cloud", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "Partial", "A812-1-EML-003"),
    ]

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey)
    sample = examples[0]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 13):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['cloud_deployment'].add(ws[f'C{r}'])

    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'K']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    # J column (Quarantine) - Yes/No/Partial/N/A
    quarantine_val = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(quarantine_val)
    for r in range(6, 56):
        quarantine_val.add(ws[f'J{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_cloud_casb_dlp(ws, styles):
    """Create Cloud/CASB DLP Assessment sheet."""

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "CLOUD / CASB DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Assess Cloud Access Security Broker and cloud-native DLP"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Cloud Service"),
        ("B4", "CASB Solution"),
        ("C4", "Integration Type"),
        ("D4", "Upload Monitoring"),
        ("E4", "Download Monitoring"),
        ("F4", "Sharing Controls"),
        ("G4", "Content Inspection"),
        ("H4", "Real-Time Blocking"),
        ("I4", "Shadow IT Discovery"),
        ("J4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [25, 25, 18, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example data
    examples = [
        ("Microsoft OneDrive", "Netskope", "API", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-CLD-001"),
        ("Dropbox", "Netskope", "API", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-CLD-002"),
        ("Box", "McAfee MVISION", "API", "Yes", "Partial", "Yes", "Yes", "Partial", "Partial", "A812-1-CLD-003"),
        ("Google Drive", "Native Google DLP", "API", "Yes", "Yes", "Yes", "Yes", "Yes", "N/A", "A812-1-CLD-004"),
        ("GitHub", "Netskope", "API", "Yes", "Yes", "No", "Yes", "Yes", "Yes", "A812-1-CLD-005"),
    ]

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey)
    sample = examples[0]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['casb_integration'].add(ws[f'C{r}'])

    for col in ['D', 'E', 'F', 'G', 'H', 'I']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_web_dlp(ws, styles):
    """Create Web DLP Assessment sheet."""

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "WEB PROXY DLP ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Assess web proxy DLP configurations (forward/reverse proxy, SSL inspection)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Proxy Solution"),
        ("B4", "Proxy Type"),
        ("C4", "SSL Inspection"),
        ("D4", "HTTP/HTTPS Coverage"),
        ("E4", "Upload Monitoring"),
        ("F4", "Download Monitoring"),
        ("G4", "Cloud Storage Detection"),
        ("H4", "Webmail Monitoring"),
        ("I4", "URL Filtering Integration"),
        ("J4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [25, 18, 18, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example data
    examples = [
        ("Zscaler", "Cloud", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-WEB-001"),
        ("Forcepoint Web", "Forward", "Yes", "Yes", "Yes", "Partial", "Yes", "Yes", "Yes", "A812-1-WEB-002"),
        ("Squid Proxy", "Forward", "No", "Partial", "No", "No", "No", "No", "Partial", "A812-1-WEB-003"),
    ]

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey)
    sample = examples[0]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['proxy_type'].add(ws[f'B{r}'])

    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    # I column (URL Filtering) - Yes/No/Partial/N/A
    url_val = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(url_val)
    for r in range(6, 56):
        url_val.add(ws[f'I{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_database_dam(ws, styles):
    """Create Database Activity Monitoring sheet."""

    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "DATABASE ACTIVITY MONITORING (DAM)"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Assess Database Activity Monitoring for DLP use cases"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Database System"),
        ("B4", "DAM Solution"),
        ("C4", "Monitoring Scope"),
        ("D4", "SELECT Query Monitoring"),
        ("E4", "Bulk Export Detection"),
        ("F4", "Privileged User Monitoring"),
        ("G4", "DLP Policy Integration"),
        ("H4", "Alert Triggering"),
        ("I4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [25, 25, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example data
    examples = [
        ("Oracle xxx", "Imperva DAM", "Production", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-DAM-001"),
        ("SQL Server xxxx", "IBM Guardium", "Production", "Yes", "Yes", "Yes", "Yes", "Yes", "A812-1-DAM-002"),
        ("MySQL x.x", "Native Audit Log", "All", "Partial", "No", "Partial", "No", "Partial", "A812-1-DAM-003"),
    ]

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey)
    sample = examples[0]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['db_scope'].add(ws[f'C{r}'])

    for col in ['D', 'E', 'F', 'G', 'H']:
        for r in range(6, 56):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: SHEET CREATION FUNCTIONS - PART 3 (STANDARD SHEETS)
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet (standard across all workbooks)."""

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "GAP ANALYSIS - IDENTIFIED DEFICIENCIES"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Document all infrastructure gaps and remediation plans"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: EMPTY spacing row

    # Column headers at row 4
    headers = [
        ("A4", "Gap ID"),
        ("B4", "Gap Description"),
        ("C4", "Affected Technology"),
        ("D4", "Risk Level"),
        ("E4", "Business Impact"),
        ("F4", "Root Cause"),
        ("G4", "Remediation Plan"),
        ("H4", "Owner"),
        ("I4", "Target Date"),
        ("J4", "Status"),
        ("K4", "Evidence ID"),
    ]

    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])

    # Column widths
    widths = [18, 35, 25, 15, 30, 30, 35, 20, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # All example gap data
    examples = [
        ("GAP-A812-1-001", "Endpoint DLP not deployed to Linux servers", "Endpoint DLP", "High", "Unmonitored data exfiltration from Linux systems", "Budget constraints, vendor compatibility", "Deploy compatible endpoint agent or implement compensating controls", "[Owner]", "[Date]", "Open", "A812-1-GAP-001"),
        ("GAP-A812-1-002", "No SSL inspection on web proxy", "Web DLP", "High", "Encrypted channel bypass for web uploads", "Privacy concerns, certificate management complexity", "Implement SSL inspection with privacy impact assessment", "[Owner]", "[Date]", "Open", "A812-1-GAP-002"),
        ("GAP-A812-1-003", "CASB missing for Dropbox", "Cloud DLP", "Medium", "Shadow IT data leakage via Dropbox", "Unknown usage, not in approved SaaS list", "Deploy CASB coverage or block Dropbox entirely", "[Owner]", "[Date]", "Open", "A812-1-GAP-003"),
    ]

    # Row 5: Sample row — ALL columns populated (F2F2F2 grey)
    sample = examples[0]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6 (after sample row 5)
    for r in range(6, 56):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])

    # Data validations (start at row 6, after sample row 5)
    validations = create_data_validations()

    for r in range(6, 56):
        validations['risk_level'].add(ws[f'D{r}'])

    for r in range(6, 56):
        validations['gap_status'].add(ws[f'J{r}'])

    ws.freeze_panes = 'A5'
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete each worksheet tab in sequence.',
        '2. Fill all yellow-highlighted cells with your organisation\'s information.',
        '3. Use dropdown menus where provided.',
        '4. Document all DLP technologies across network, endpoint, cloud, email, web, and database layers.',
        '5. Provide evidence references for all assessments.',
        '6. Review Summary Dashboard for overall compliance score.',
        '7. Identify gaps in Gap Analysis sheet and create remediation plans.',
        '8. Maintain the Evidence Register for audit traceability.',
        '9. Obtain final approval in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A23"] = "Status Legend"
    ws["A23"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=24, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 25 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet matching gold standard."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 9):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Evidence ID", "Assessment Area", "Evidence Type", "Description",
               "Location/Path", "Date Collected", "Collected By", "Verification Status"]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C6:C105")

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_ver)
    dv_ver.add("H6:H105")

    # Row 5: F2F2F2 sample row
    INFO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = {
        1: "EV-001", 2: "DLP Technology Inventory", 3: "Configuration file",
        4: "DLP policy rule set configuration export",
        5: "\\\\fileserver\\isms\\evidence\\dlp\\policy-rules.pdf",
        6: "22.02.2026", 7: "IT Security Team", 8: "Verified"
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = INFO_FILL
        cell.border = border

    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    ws.freeze_panes = "A5"


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet matching gold standard."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO/IEC 27001:2022 \u2014 Control A.8.12: {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    summary_fields = [
        ("Document:", f"{WORKBOOK_ID} - {ASSESSMENT_AREA}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "='Summary Dashboard'!G14"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 2}"])

    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard sheet — Gold Standard TABLE 1/2/3 layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title (A1:G1 merged) ──────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (A2:G2 merged) ───────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.12: Data Leakage Prevention"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: Blank ──────────────────────────────────────────────────────────

    # ── TABLE 1: COMPLIANCE ASSESSMENT SUMMARY ────────────────────────────────
    # Row 4: Banner
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].border = border

    # Row 5: Column headers (D9D9D9)
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # TABLE 1 data rows — each row = one assessment area
    # Sample row is row 5 in each data sheet; COUNTIF ranges start at row 6
    # For yes_no_partial DV: Compliant="Yes", Partial="Partial", Non-Compliant="No", N/A="N/A"
    # For integration_status DV: Compliant="Integrated", Partial="Partial", Non-Compliant="Standalone"
    # For gap_status DV: Compliant="Resolved"+"Closed"+"Accepted", Partial="In Progress", Non-Compliant="Open"
    t1_data = [
        # (Area, Compliant formula, Partial formula, Non-Compliant formula, N/A formula)
        # DLP Technology Inventory — key col M (Integration Status): Integrated/Standalone/Partial
        # Sample row 5; empty rows 6-55 → range M6:M55
        ("DLP Technology Inventory",
         "=COUNTIF('DLP Technology Inventory'!M6:M55,\"Integrated\")",
         "=COUNTIF('DLP Technology Inventory'!M6:M55,\"Partial\")",
         "=COUNTIF('DLP Technology Inventory'!M6:M55,\"Standalone\")",
         "=0"),
        # Network DLP — key col H (Content Inspection): Yes/No/Partial/N/A
        # Sample row 5; empty rows 6-55 → range H6:H55
        ("Network DLP Assessment",
         "=COUNTIF('Network DLP'!H6:H55,\"Yes\")",
         "=COUNTIF('Network DLP'!H6:H55,\"Partial\")",
         "=COUNTIF('Network DLP'!H6:H55,\"No\")",
         "=COUNTIF('Network DLP'!H6:H55,\"N/A\")"),
        # Endpoint DLP — key col I (USB Blocking): Yes/No/Partial/N/A
        # Sample row 5; empty rows 6-55 → range I6:I55
        ("Endpoint DLP Assessment",
         "=COUNTIF('Endpoint DLP'!I6:I55,\"Yes\")",
         "=COUNTIF('Endpoint DLP'!I6:I55,\"Partial\")",
         "=COUNTIF('Endpoint DLP'!I6:I55,\"No\")",
         "=COUNTIF('Endpoint DLP'!I6:I55,\"N/A\")"),
        # Email DLP — key col D (SMTP Gateway Protection): Yes/No/Partial/N/A
        # Sample row 5; empty rows 6-55 → range D6:D55
        ("Email DLP Assessment",
         "=COUNTIF('Email DLP'!D6:D55,\"Yes\")",
         "=COUNTIF('Email DLP'!D6:D55,\"Partial\")",
         "=COUNTIF('Email DLP'!D6:D55,\"No\")",
         "=COUNTIF('Email DLP'!D6:D55,\"N/A\")"),
        # Cloud CASB DLP — key col G (Content Inspection): Yes/No/Partial/N/A
        # Sample row 5; empty rows 6-55 → range G6:G55
        ("Cloud / CASB DLP Assessment",
         "=COUNTIF('Cloud CASB DLP'!G6:G55,\"Yes\")",
         "=COUNTIF('Cloud CASB DLP'!G6:G55,\"Partial\")",
         "=COUNTIF('Cloud CASB DLP'!G6:G55,\"No\")",
         "=COUNTIF('Cloud CASB DLP'!G6:G55,\"N/A\")"),
        # Web DLP — key col C (SSL Inspection): Yes/No/Partial/N/A
        # Sample row 5; empty rows 6-55 → range C6:C55
        ("Web DLP Assessment",
         "=COUNTIF('Web DLP'!C6:C55,\"Yes\")",
         "=COUNTIF('Web DLP'!C6:C55,\"Partial\")",
         "=COUNTIF('Web DLP'!C6:C55,\"No\")",
         "=COUNTIF('Web DLP'!C6:C55,\"N/A\")"),
        # Database DAM — key col D (SELECT Query Monitoring): Yes/No/Partial/N/A
        # Sample row 5; empty rows 6-55 → range D6:D55
        ("Database Activity Monitoring",
         "=COUNTIF('Database DAM'!D6:D55,\"Yes\")",
         "=COUNTIF('Database DAM'!D6:D55,\"Partial\")",
         "=COUNTIF('Database DAM'!D6:D55,\"No\")",
         "=COUNTIF('Database DAM'!D6:D55,\"N/A\")"),
        # Gap Analysis — key col J (Status): Resolved+Closed+Accepted vs In Progress vs Open
        # Sample row 5; empty rows 6-55 → range J6:J55
        ("Gap Analysis",
         "=COUNTIF('Gap Analysis'!J6:J55,\"Resolved\")+COUNTIF('Gap Analysis'!J6:J55,\"Closed\")+COUNTIF('Gap Analysis'!J6:J55,\"Accepted\")",
         "=COUNTIF('Gap Analysis'!J6:J55,\"In Progress\")",
         "=COUNTIF('Gap Analysis'!J6:J55,\"Open\")",
         "=0"),
    ]

    t1_start = 6
    blue_font = Font(name="Calibri", size=10, color="000000")
    for i, (area, compl, partial, non_compl, na) in enumerate(t1_data):
        r = t1_start + i
        ws.cell(row=r, column=1, value=area).border = border
        ws.cell(row=r, column=1).font = blue_font
        # Col B = SUM(C+D+E+F)
        ws.cell(row=r, column=2, value=f"=C{r}+D{r}+E{r}+F{r}").border = border
        ws.cell(row=r, column=2).font = blue_font
        ws.cell(row=r, column=3, value=compl).border = border
        ws.cell(row=r, column=3).font = blue_font
        ws.cell(row=r, column=4, value=partial).border = border
        ws.cell(row=r, column=4).font = blue_font
        ws.cell(row=r, column=5, value=non_compl).border = border
        ws.cell(row=r, column=5).font = blue_font
        ws.cell(row=r, column=6, value=na).border = border
        ws.cell(row=r, column=6).font = blue_font
        # Col G = Compliance %: IF((B-F)=0,0,C/(B-F))
        ws.cell(row=r, column=7, value=f"=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))").border = border
        ws.cell(row=r, column=7).font = blue_font
        ws.cell(row=r, column=7).number_format = "0.0%"

    # TOTAL row
    t1_total_row = t1_start + len(t1_data)
    ws.cell(row=t1_total_row, column=1, value="TOTAL").border = border
    ws.cell(row=t1_total_row, column=1).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=t1_total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col in range(2, 7):
        cell = ws.cell(row=t1_total_row, column=col)
        cell.border = border
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", bold=True, size=10)
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}{t1_start}:{col_letter}{t1_total_row - 1})"
    # TOTAL Compliance %
    ws.cell(row=t1_total_row, column=7,
            value=f"=IF((B{t1_total_row}-F{t1_total_row})=0,0,C{t1_total_row}/(B{t1_total_row}-F{t1_total_row}))").border = border
    ws.cell(row=t1_total_row, column=7).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=t1_total_row, column=7).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=t1_total_row, column=7).number_format = "0.0%"

    # Blank row after TABLE 1
    current_row = t1_total_row + 2

    # ── TABLE 2: KEY METRICS ──────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:G{current_row}")
    ws[f"A{current_row}"] = "TABLE 2 \u2013 KEY METRICS"
    ws[f"A{current_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{current_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{current_row}"].border = border
    current_row += 1

    # TABLE 2 header row
    t2_headers = ["Metric", "Value", "Target", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=hdr if hdr else "")
        if hdr:
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    current_row += 1

    t2_data = [
        ("DLP Technologies in Production",
         "=COUNTIF('DLP Technology Inventory'!G6:G55,\"Production\")",
         "\u22654"),
        ("SIEM Integration Rate",
         "=IF(COUNTA('DLP Technology Inventory'!B6:B55)=0,0,COUNTIF('DLP Technology Inventory'!N6:N55,\"Yes\")/COUNTA('DLP Technology Inventory'!B6:B55))",
         "\u226590%"),
        ("Endpoint DLP Coverage (avg %)",
         "=IFERROR(AVERAGE('Endpoint DLP'!H6:H55),0)",
         "\u226595%"),
        ("Open Critical/High Gaps",
         "=COUNTIFS('Gap Analysis'!D6:D55,\"Critical\",'Gap Analysis'!J6:J55,\"Open\")+COUNTIFS('Gap Analysis'!D6:D55,\"High\",'Gap Analysis'!J6:J55,\"Open\")",
         "0"),
        ("Gap Remediation Rate",
         "=IF(COUNTA('Gap Analysis'!A6:A55)=0,0,(COUNTIF('Gap Analysis'!J6:J55,\"Resolved\")+COUNTIF('Gap Analysis'!J6:J55,\"Closed\"))/COUNTA('Gap Analysis'!A6:A55))",
         "\u226580%"),
        ("Evidence Items Verified",
         "=COUNTIF('Evidence Register'!H5:H104,\"Verified\")",
         "\u226580%"),
    ]

    for metric, value_formula, target in t2_data:
        ws.cell(row=current_row, column=1, value=metric).border = border
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=current_row, column=2, value=value_formula).border = border
        ws.cell(row=current_row, column=2).font = Font(name="Calibri", size=10)
        ws.cell(row=current_row, column=3, value=target).border = border
        ws.cell(row=current_row, column=3).font = Font(name="Calibri", size=10)
        for col in range(4, 8):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1

    current_row += 1  # blank row

    # ── TABLE 3: KEY FINDINGS & RECOMMENDATIONS ───────────────────────────────
    ws.merge_cells(f"A{current_row}:G{current_row}")
    ws[f"A{current_row}"] = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    ws[f"A{current_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{current_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{current_row}"].border = border
    current_row += 1

    # TABLE 3 header row
    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", ""]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=hdr if hdr else "")
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    current_row += 1

    # 5 FFFFCC finding rows
    for i in range(1, 6):
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
        ws.cell(row=current_row, column=1).value = str(i)
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=10)
        current_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    
    logger.info("=" * 78)
    logger.info(f"{WORKBOOK_ID} - {ASSESSMENT_AREA} Generator")
    logger.info(f"ISO/IEC 27001:2022 Control {CONTROL_ID}")
    logger.info("=" * 78)
    
    wb = create_workbook()
    styles = _STYLES
    
    logger.info("\n[1/12] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("[2/12] Creating DLP Technology Inventory...")
    create_technology_inventory(wb["DLP Technology Inventory"], styles)

    logger.info("[3/12] Creating Network DLP Assessment...")
    create_network_dlp(wb["Network DLP"], styles)

    logger.info("[4/12] Creating Endpoint DLP Assessment...")
    create_endpoint_dlp(wb["Endpoint DLP"], styles)

    logger.info("[5/12] Creating Email DLP Assessment...")
    create_email_dlp(wb["Email DLP"], styles)

    logger.info("[6/12] Creating Cloud/CASB DLP Assessment...")
    create_cloud_casb_dlp(wb["Cloud CASB DLP"], styles)

    logger.info("[7/12] Creating Web DLP Assessment...")
    create_web_dlp(wb["Web DLP"], styles)

    logger.info("[8/12] Creating Database DAM Assessment...")
    create_database_dam(wb["Database DAM"], styles)

    logger.info("[9/12] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[10/12] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[11/12] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[12/12] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.12.1_DLP_Infrastructure_{datetime.now().strftime('%Y%m%d')}.xlsx"
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  \u2022 Instructions & Legend")
    logger.info("  \u2022 DLP Technology Inventory (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Network DLP (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Endpoint DLP (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Email DLP (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Cloud/CASB DLP (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Web DLP (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Database DAM (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Gap Analysis (51 rows: 1 sample + 50 empty)")
    logger.info("  \u2022 Summary Dashboard (12 KPIs + compliance tracking)")
    logger.info("  \u2022 Evidence Register (100 rows)")
    logger.info("  \u2022 Approval Sign-Off")
    logger.info("\n" + "=" * 78)
    logger.info("NEXT STEPS:")
    logger.info("1. Open the workbook in Excel/LibreOffice")
    logger.info("2. Complete yellow-highlighted cells")
    logger.info("3. Use dropdowns (don't type free-form in dropdown cells)")
    logger.info("4. Gather evidence for all assessments")
    logger.info("5. Review Summary Dashboard for compliance score")
    logger.info("6. Obtain CISO/DPO approval")
    logger.info("=" * 78)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
