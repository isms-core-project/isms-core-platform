#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.4 - Monitoring & Response Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 4 of 4: Monitoring & Response

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific monitoring capabilities, incident response
procedures, and SOC operations.

Key customization areas:
1. SIEM platform and integration (specific to your technology stack)
2. Alert thresholds and SLAs (aligned with your operations)
3. Incident response workflows (per your IR procedures)
4. SOC escalation procedures (based on your organisational structure)
5. Dashboard and reporting requirements (per stakeholder needs)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates comprehensive monitoring and response assessment workbook for systematic
evaluation of DLP detection, alerting, and incident response capabilities against
ISO 27001:2022 Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• Logging configuration for all DLP channels
• Alert rules inventory and alert volume metrics
• SIEM integration and log correlation
• False positive management and tuning procedures
• Incident response workflow and SLA compliance
• SOC integration and escalation procedures
• Executive and operational dashboards
• Gap analysis and remediation planning
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.4_Monitoring_Response_YYYYMMDD.xlsx

Sheets (13 total):
1. Instructions & Legend - Assessment guidance and metadata
2. Logging Configuration - DLP logging requirements per channel
3. Alert Rules Inventory - All configured DLP alert rules
4. Alert Volume Metrics - Daily/weekly/monthly alert statistics
5. SIEM Integration - Log forwarding, correlation rules, use cases
6. False Positive Management - FP rate tracking, tuning procedures
7. Incident Response Workflow - Triage, containment, investigation
8. SOC Integration - Alert workflow, escalation, SLAs
9. Dashboards Reporting - Executive, operational, compliance dashboards
10. Gap Analysis - Monitoring and response gaps (40 rows)
11. Summary Dashboard - Compliance metrics and KPIs
12. Evidence Register - Audit evidence tracking (100 rows)
13. Approval Sign-Off - 3-level assessment approval workflow

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_4_monitoring_response.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.4_Monitoring_Response_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organisation-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Document all DLP logging configurations
    5. Inventory alert rules and analyse alert volumes
    6. Assess SIEM integration and correlation capabilities
    7. Measure false positive rates and tuning effectiveness
    8. Verify incident response procedures and SLA compliance
    9. Collect and document evidence (Evidence Register sheet)
    10. Complete gap analysis for monitoring deficiencies
    11. Obtain management approval (Summary Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 4 of 4 in comprehensive DLP evaluation framework
    Focus: Detection, monitoring, alerting, and incident response
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.3 (Monitoring & Detection Requirements)
    Policy:         ISMS-POL-A.8.12-S2.4 (Incident Response & Remediation)
    Policy:         ISMS-POL-A.8.12-S5.C (DLP Incident Response Procedures)
    Implementation: ISMS-IMP-A.8.12.4 (Monitoring & Response Implementation Guide)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py
       python3 generate_a812_2_data_classification.py
       python3 generate_a812_3_channel_coverage.py
       python3 generate_a812_4_monitoring_response.py      ← YOU ARE HERE
    
    2. Complete assessments (manual - security team, SOC analysts)
    
    3. Normalise filenames for dashboard linking:
       python3 normalise_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Monitoring Assessment → Normalise → Dashboard → Audit Evidence

Critical Prerequisites:
    • Domain 1 (Infrastructure) - need DLP technology capabilities
    • Domain 3 (Channel Coverage) - need alert rules per channel
    • SIEM/SOC team involvement for integration assessment

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    4 of 4 (Monitoring & Response)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organisation] ISMS Implementation Team
License:              [Organisation License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - ISO/IEC 27035:2023 (Information Security Incident Management)
    - Swiss FADP (Federal Act on Data Protection - Breach Notification)
    - EU GDPR (General Data Protection Regulation - Articles 33, 34)
    - NIST SP 800-53 (Security and Privacy Controls - IR, SI families)
    - NIST SP 800-61 (Computer Security Incident Handling Guide)
    - CIS Controls v8.1 (Center for Internet Security - Controls 8, 17)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Alert Severity Classification (from ISMS-POL-A.8.12-S2.3):
    Critical (Response SLA: 15 minutes):
        • Active data exfiltration detected
        • Credentials (passwords, API keys) detected in transit
        • Large-volume sensitive data transfers
    
    High (Response SLA: 1 hour):
        • PII/IP transferred externally without encryption
        • Sensitive data sent to unauthorised cloud services
        • Policy violations by privileged users
    
    Medium (Response SLA: 4 hours):
        • Confidential data transfers requiring review
        • Repeated policy violations by single user
        • Suspicious patterns requiring investigation
    
    Low (Response SLA: 24 hours):
        • Policy violations requiring tuning
        • User education opportunities
        • Process improvement notifications
    
    Informational (No SLA):
        • Monitoring-only events
        • Audit trail entries
        • Compliance reporting data

Swiss FADP Breach Notification Requirements:
    Data breach notification obligations under Swiss Federal Act on Data Protection:
    • Notification to Federal Data Protection and Information Commissioner (FDPIC)
    • Notification required when breach likely to result in high risk to data subjects
    • Notification must be "without delay" (typically within 72 hours)
    • DLP monitoring must detect breaches in time to meet notification deadlines
    
    This assessment verifies:
    • DLP alerts configured to detect breach scenarios
    • Incident response procedures align with FADP timelines
    • Evidence collection supports breach notification reporting

Assessment Scope - Monitoring Components:
    1. Logging Configuration:
       • Log all DLP events (alerts, policy matches, data transfers)
       • Retention period: minimum 90 days (recommend 1 year)
       • Log format: structured (JSON/CEF) for SIEM ingestion
       • Protected logs: tamper-proof, encrypted storage
    
    2. Alert Rules:
       • Content inspection rules (keywords, regex patterns)
       • Data classification label rules
       • File type and size threshold rules
       • User/group exception rules
       • Contextual rules (time, location, device)
    
    3. SIEM Integration:
       • Real-time log forwarding (syslog, API)
       • Correlation rules (DLP + network + endpoint)
       • Use cases (insider threat, data exfiltration)
       • Alert enrichment (user context, asset context)
    
    4. False Positive Management:
       • FP rate measurement (target: <5%)
       • Root cause analysis procedures
       • Policy tuning workflows
       • Exception management process
    
    5. Incident Response:
       • Triage procedures (severity classification)
       • Containment actions (block user, quarantine file)
       • Investigation steps (forensics, user interview)
       • Remediation and lessons learned

Key Performance Indicators (KPIs):
    • Alert Volume: Daily/weekly/monthly trends
    • False Positive Rate: Target <5%
    • Mean Time to Detect (MTTD): Target <1 hour for Critical
    • Mean Time to Respond (MTTR): Per severity SLA
    • Incident Closure Rate: % within SLA
    • Tuning Effectiveness: Reduction in FP rate over time

Data Classification:
    Generated workbooks contain sensitive organisational security information.
    Handle according to [Organisation]'s data classification policy.
    Recommended classification: [Organisation] Internal/Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: logging completeness, SIEM integration, incident response.

Review Cycle:
    Monthly: Review alert volumes and false positive rates
    Quarterly: Update monitoring configuration and tuning status
    Annually: Complete full monitoring and response reassessment
    Ad-hoc: After security incidents or infrastructure changes

================================================================================
--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.4"
WORKBOOK_NAME = "Monitoring & Response Assessment"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.4"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.3, ISMS-POL-A.8.12-S2.4"
ASSESSMENT_AREA = "Monitoring & Response Assessment"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "F2F2F2"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "F2F2F2"         # Light blue (Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 40
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create all sheets in order
    sheets = [
        "Instructions & Legend",
        "Logging Configuration",
        "Alert Rules Inventory",
        "Alert Volume Metrics",
        "SIEM Integration",
        "False Positive Management",
        "Incident Response Workflow",
        "SOC Integration",
        "Dashboards Reporting",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    return {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "font": Font(name="Calibri", size=10, italic=True),
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "data_cell": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
    }



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style template to cell by creating NEW instances.
    
    CRITICAL: Each cell gets its own Font/Fill/Border/Alignment objects.
    This prevents openpyxl's "Cannot use already used style object" error.
    """
    if "font" in style_dict:
        f = style_dict["font"]
        cell.font = Font(
            name=f.name,
            size=f.size,
            bold=f.bold,
            italic=f.italic,
            color=f.color
        )
    
    if "fill" in style_dict:
        f = style_dict["fill"]
        cell.fill = PatternFill(
            start_color=f.start_color.rgb if f.start_color else None,
            end_color=f.end_color.rgb if f.end_color else None,
            fill_type=f.fill_type
        )
    
    if "alignment" in style_dict:
        a = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=a.horizontal,
            vertical=a.vertical,
            wrap_text=a.wrap_text
        )
    
    if "border" in style_dict:
        b = style_dict["border"]
        cell.border = Border(
            left=Side(style=b.left.style) if b.left else None,
            right=Side(style=b.right.style) if b.right else None,
            top=Side(style=b.top.style) if b.top else None,
            bottom=Side(style=b.bottom.style) if b.bottom else None
        )


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create fresh data validation objects for a single worksheet."""
    
    # Standard response values (Yes/No/Partial/Planned/N/A)
    val_response = DataValidation(
        type="list",
        formula1='"Yes,No,Partial,Planned,N/A"',
        allow_blank=False
    )
    
    # Alert severity levels
    val_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,Informational"',
        allow_blank=False
    )
    
    # Channel types
    val_channel = DataValidation(
        type="list",
        formula1='"Email,Web,Endpoint,Network,Application,Mobile"',
        allow_blank=False
    )
    
    # Detection methods
    val_detection = DataValidation(
        type="list",
        formula1='"Pattern,Keyword,Fingerprint,Contextual,ML"',
        allow_blank=False
    )
    
    # Policy actions
    val_action = DataValidation(
        type="list",
        formula1='"Allow,Alert,Block,Quarantine,Encrypt"',
        allow_blank=False
    )
    
    # Alert destinations
    val_destination = DataValidation(
        type="list",
        formula1='"SIEM,Email,SMS,Ticketing,Dashboard"',
        allow_blank=False
    )
    
    # Response SLAs
    val_sla = DataValidation(
        type="list",
        formula1='"15min,1hr,4hr,24hr,No SLA"',
        allow_blank=False
    )
    
    # Rule status
    val_rule_status = DataValidation(
        type="list",
        formula1='"Active,Disabled,Testing,Deprecated"',
        allow_blank=False
    )
    
    # Evidence types
    val_evidence = DataValidation(
        type="list",
        formula1='"Config,Screenshot,Log,Report,Dashboard,Other"',
        allow_blank=False
    )
    
    # Verification status
    val_verification = DataValidation(
        type="list",
        formula1='"Verified,Pending,Rejected"',
        allow_blank=False
    )
    
    # Risk levels
    val_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    
    # Gap status
    val_gap_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,Blocked"',
        allow_blank=False
    )
    
    # Affected areas for gaps
    val_affected_area = DataValidation(
        type="list",
        formula1='"Logging,Alerting,SIEM,FP Management,Incident Response,SOC,Dashboards"',
        allow_blank=False
    )
    
    return {
        "response": val_response,
        "severity": val_severity,
        "channel": val_channel,
        "detection": val_detection,
        "action": val_action,
        "destination": val_destination,
        "sla": val_sla,
        "rule_status": val_rule_status,
        "evidence": val_evidence,
        "verification": val_verification,
        "risk": val_risk,
        "gap_status": val_gap_status,
        "affected_area": val_affected_area,
    }


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 1
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete Logging Configuration first — this is the foundational data layer.', '2. Document all Alert Rules in the Inventory (what alerts exist, thresholds, actions).', '3. Track Alert Volume Metrics (daily/weekly/monthly statistics by severity and channel).', '4. Verify SIEM Integration (log forwarding, correlation rules, use cases).', '5. Measure False Positive rates and document tuning procedures.', '6. Document the Incident Response Workflow (triage, containment, SLA compliance).', '7. Assess SOC Integration (alert workflow, escalation, staffing adequacy).', '8. Verify Dashboards and Reporting capabilities (executive, operational, compliance).', '9. Identify gaps and create remediation plans in the Gap Analysis sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_logging_configuration_sheet(ws, styles):
    """Create Logging Configuration sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "DLP LOGGING CONFIGURATION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess DLP logging requirements per channel (25 criteria)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Assessment criteria
    requirements = [
        "Are all DLP channels configured to log events?",
        "Email DLP: Logs sent/received/blocked emails with metadata?",
        "Web DLP: Logs uploads/downloads/blocked web traffic?",
        "Endpoint DLP: Logs USB/print/screen capture/clipboard events?",
        "Network DLP: Logs file transfers (FTP/SMB/SCP)?",
        "Application DLP: Logs database exports/API calls?",
        "Mobile DLP: Logs mobile app usage and data transfers?",
        "Do logs include user identity (username, email, employee ID)?",
        "Do logs include timestamp (ISO 8601 format)?",
        "Do logs include source (IP, hostname, device ID)?",
        "Do logs include destination (recipient, URL, file path)?",
        "Do logs include data classification level (if available)?",
        "Do logs include policy violated (rule ID, rule name)?",
        "Do logs include action taken (allow/alert/block/quarantine)?",
        "Do logs include file metadata (filename, size, hash)?",
        "Are logs retained for minimum 12 months (FADP requirement)?",
        "Are logs protected from unauthorised modification?",
        "Are logs encrypted in transit (TLS to SIEM)?",
        "Are logs encrypted at rest (SIEM storage)?",
        "Is log integrity verified (checksums, digital signatures)?",
        "Are logs reviewed for completeness (missing log gaps)?",
        "Are logs backed up (separate from production logs)?",
        "Is log access restricted (RBAC, audit trail of log access)?",
        "Are logs parsed correctly by SIEM (no parsing errors)?",
        "Are logs correlated with other security events (IAM, EDR, Network)?",
    ]

    # Row 5: sample row (first grey example — all cols F2F2F2)
    ws['A5'] = 1
    ws['B5'] = requirements[0]
    ws['C5'] = "Yes"
    ws['D5'] = "A812-4-LOG-001"
    ws['E5'] = "All channels logging to Splunk"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}5'], styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    for row in range(6, 56):
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])

    # Apply data validation to Compliance column — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["response"].add(ws[f'C{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE

    # Freeze panes
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A5'


def create_alert_rules_inventory_sheet(ws, styles):
    """Create Alert Rules Inventory sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "DLP ALERT RULES INVENTORY"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:M2')
    ws['A2'] = "Complete inventory of all configured DLP alert rules (30 rows)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = [
        "Alert Rule ID", "Rule Name", "Channel", "Severity", "Detection Method",
        "Policy Action", "Alert Destination", "Response SLA", "Last Tuned Date",
        "FP Rate %", "Alert Volume (30d)", "Rule Status", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Row 5: sample row (first grey example — all cols F2F2F2)
    sample_rule = ("ALT-001", "Credentials Detected - Email", "Email", "Critical", "Pattern",
                   "Block", "SIEM", "15min", "2025-01-01", 2.1, 45, "Active", "A812-4-ALT-001")
    for col_idx, value in enumerate(sample_rule, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    for row in range(6, 56):
        for col_idx in range(1, 14):
            apply_style(ws.cell(row=row, column=col_idx), styles["input_cell"])

    # Apply data validations — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["channel"].add(ws[f'C{r}'])
        validations["severity"].add(ws[f'D{r}'])
        validations["detection"].add(ws[f'E{r}'])
        validations["action"].add(ws[f'F{r}'])
        validations["destination"].add(ws[f'G{r}'])
        validations["sla"].add(ws[f'H{r}'])
        validations["rule_status"].add(ws[f'L{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_NARROW
    ws.column_dimensions['I'].width = WIDTH_MEDIUM
    ws.column_dimensions['J'].width = WIDTH_NARROW
    ws.column_dimensions['K'].width = WIDTH_MEDIUM
    ws.column_dimensions['L'].width = WIDTH_MEDIUM
    ws.column_dimensions['M'].width = WIDTH_MEDIUM

    # Freeze panes
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 2
# ============================================================================

def create_alert_volume_metrics_sheet(ws, styles):
    """Create Alert Volume Metrics sheet."""

    # Header
    ws.merge_cells('A1:G1')
    ws['A1'] = "DLP ALERT VOLUME METRICS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:G2')
    ws['A2'] = "Track alert volume by severity and channel (Last 30 Days)"
    apply_style(ws['A2'], styles["subheader"])

    row = 4
    
    # Section 1: Alert Volume by Severity
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "ALERT VOLUME BY SEVERITY (Last 30 Days)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    severity_headers = ["Severity", "Total Alerts", "Avg per Day", "Blocked", "Alerted", "False Positives", "FP Rate %"]
    for col_idx, header in enumerate(severity_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    severities = ["Critical", "High", "Medium", "Low", "Informational", "TOTAL"]
    for sev in severities:
        ws[f'A{row}'] = sev
        if sev == "TOTAL":
            ws[f'A{row}'].font = Font(bold=True)
            # Formulas for TOTAL row
            ws[f'B{row}'] = f'=SUM(B{row-5}:B{row-1})'
            ws[f'C{row}'] = f'=B{row}/30'
            ws[f'D{row}'] = f'=SUM(D{row-5}:D{row-1})'
            ws[f'E{row}'] = f'=SUM(E{row-5}:E{row-1})'
            ws[f'F{row}'] = f'=SUM(F{row-5}:F{row-1})'
            ws[f'G{row}'] = f'=IF(B{row}>0,F{row}/B{row}*100,0)'
        else:
            # Input cells for data entry
            for col in ['B', 'C', 'D', 'E', 'F']:
                apply_style(ws[f'{col}{row}'], styles["input_cell"])
            # FP Rate formula
            ws[f'G{row}'] = f'=IF(B{row}>0,F{row}/B{row}*100,0)'
        row += 1
    
    row += 1
    
    # Section 2: Alert Volume by Channel
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ALERT VOLUME BY CHANNEL (Last 30 Days)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    channel_headers = ["Channel", "Total Alerts", "Blocked", "Alerted", "FP Rate %", "Top Rule Name"]
    for col_idx, header in enumerate(channel_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    channels = ["Email", "Web", "Endpoint", "Network", "Application", "Mobile"]
    for channel in channels:
        ws[f'A{row}'] = channel
        for col in ['B', 'C', 'D', 'E', 'F']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])
        row += 1
    
    row += 1
    
    # Section 3: Alert Trend Analysis (4 weeks)
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "ALERT TREND ANALYSIS (Weekly)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    trend_headers = ["Week", "Critical", "High", "Medium", "Low", "Total", "FP Rate %", "Notes"]
    for col_idx, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    weeks = ["Week 1 (Dec 2-8)", "Week 2 (Dec 9-15)", "Week 3 (Dec 16-22)", "Week 4 (Dec 23-29)"]
    for week in weeks:
        ws[f'A{row}'] = week
        # Input cells
        for col in ['B', 'C', 'D', 'E', 'G', 'H']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])
        # Total formula
        ws[f'F{row}'] = f'=SUM(B{row}:E{row})'
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_siem_integration_sheet(ws, styles):
    """Create SIEM Integration sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "SIEM INTEGRATION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess SIEM integration, log forwarding, and correlation rules (25 criteria)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Assessment criteria
    requirements = [
        "Are DLP logs forwarded to organisational SIEM?",
        "Is log forwarding encrypted (TLS/SSL)?",
        "Is log forwarding reliable (no dropped logs >1%)?",
        "Are DLP logs parsed correctly in SIEM (no parsing errors)?",
        "Are DLP logs normalised (common schema: CEF, Syslog, JSON)?",
        "Are DLP events indexed for search and analysis?",
        "Is log retention in SIEM ≥12 months?",
        "Are SIEM correlation rules configured for DLP events?",
        "Correlation: DLP + Malware detection = compromised endpoint?",
        "Correlation: DLP + Failed logins = credential theft?",
        "Correlation: DLP + VPN from foreign country = account compromise?",
        "Correlation: DLP + HR termination record = insider threat?",
        "Are SIEM use cases documented (insider threat, compromised creds)?",
        "Are SIEM dashboards configured for DLP (real-time alert queue)?",
        "Are SIEM alerts integrated with ticketing system (ServiceNow, Jira)?",
        "Are DLP Critical alerts automatically escalated to SOC?",
        "Are DLP High alerts automatically assigned to analysts?",
        "Is DLP log ingestion monitored (alerts for log failures)?",
        "Is SIEM performance adequate (query response time <10 sec)?",
        "Are SIEM search queries documented (common investigations)?",
        "Is SIEM access restricted (RBAC, SOC analysts only)?",
        "Are SIEM searches audited (who searched for what)?",
        "Is SIEM backup configured (disaster recovery)?",
        "Are SIEM correlation rules tested (validation, false positive rate)?",
        "Are SIEM correlation rules tuned quarterly?",
    ]

    # Row 5: sample row (first grey example — all cols F2F2F2)
    ws['A5'] = 1
    ws['B5'] = requirements[0]
    ws['C5'] = "Yes"
    ws['D5'] = "A812-4-SIEM-001"
    ws['E5'] = "Forwarding to Splunk via TLS"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}5'], styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    for row in range(6, 56):
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])

    # Apply data validation — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["response"].add(ws[f'C{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE

    # Freeze panes
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A5'


def create_false_positive_management_sheet(ws, styles):
    """Create False Positive Management sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "FALSE POSITIVE MANAGEMENT ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess false positive tracking, tuning, and management (20 criteria + FP rate table)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Assessment criteria
    requirements = [
        "Is false positive rate measured per alert rule?",
        "Is overall FP rate <10% (target from policy)?",
        "Are false positives tracked in ticketing system?",
        "Is FP rate trended over time (monthly)?",
        "Are users able to report false positives (self-service portal)?",
        "Are false positive reports triaged within 24 hours?",
        "Are false positives investigated (root cause analysis)?",
        "Are DLP rules tuned based on FP analysis?",
        "Are legitimate business flows whitelisted (permanent exceptions)?",
        "Are whitelists documented (justification, approver, review date)?",
        "Are whitelists reviewed annually (remove stale exceptions)?",
        "Are high-FP rules disabled or re-tuned?",
        "Are users educated when FP is user error (not DLP issue)?",
        "Is FP tuning tracked (before/after FP rate improvement)?",
        "Are tuning changes tested before production deployment?",
        "Are tuning changes documented (change request, approval)?",
        "Are tuning changes rolled back if new issues arise?",
        "Is FP rate included in SOC performance metrics?",
        "Are FP trends reported to management (monthly)?",
        "Are FP reduction targets set and tracked?",
    ]

    # Row 5: sample row (first grey example — all cols F2F2F2)
    ws['A5'] = 1
    ws['B5'] = requirements[0]
    ws['C5'] = "Yes"
    ws['D5'] = "A812-4-FP-001"
    ws['E5'] = "FP rate tracked per rule in dashboard"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}5'], styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    row = 6
    for row in range(6, 56):
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])

    # Apply data validation — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["response"].add(ws[f'C{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    row = 57  # secondary table starts at row 57 (gap row 56)

    # FP Rate Calculation Table
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "FALSE POSITIVE RATE BY ALERT RULE (Top 10 Rules)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    fp_headers = ["Alert Rule ID", "Rule Name", "Total Alerts (30d)", "False Positives", "FP Rate %", "Target FP %", "Status"]
    for col_idx, header in enumerate(fp_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    # Example rows
    fp_examples = [
        ("ALT-001", "Credentials - Email", 45, 1, "=D{}/C{}", "5%", "\u2705 On Target"),
        ("ALT-002", "PII to External", 230, 20, "=D{}/C{}", "10%", "\u2705 On Target"),
        ("ALT-003", "Source Code GitHub", 12, 0, "=D{}/C{}", "5%", "\u2705 On Target"),
    ]
    
    for rule_id, rule_name, total, fp, formula, target, status in fp_examples:
        ws[f'A{row}'] = rule_id
        ws[f'B{row}'] = rule_name
        ws[f'C{row}'] = total
        ws[f'D{row}'] = fp
        ws[f'E{row}'] = formula.format(row, row)
        ws[f'F{row}'] = target
        ws[f'G{row}'] = status
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Blank rows for additional entries
    for _ in range(7):
        for col in ['A', 'B', 'C', 'D', 'F', 'G']:
            apply_style(ws.cell(row=row, column=ord(col)-ord('A')+1), styles["input_cell"])
        ws[f'E{row}'] = f'=IF(C{row}>0,D{row}/C{row}*100,0)'
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM

    # Freeze panes
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A5'


def create_incident_response_workflow_sheet(ws, styles):
    """Create Incident Response Workflow sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "INCIDENT RESPONSE WORKFLOW ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess incident response procedures and SLA compliance (25 criteria + SLA tracking)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Assessment criteria
    requirements = [
        "Are DLP alerts integrated with incident response workflow?",
        "Is incident response procedure documented (playbook)?",
        "Critical alerts: Are alerts triaged within 15 minutes?",
        "High alerts: Are alerts triaged within 1 hour?",
        "Medium alerts: Are alerts triaged within 4 hours?",
        "Low alerts: Are alerts triaged within 24 hours?",
        "Is alert triage documented (who, when, decision)?",
        "Are false positives closed with justification?",
        "Are true positives escalated to incident response team?",
        "Are users contacted for business justification (High alerts)?",
        "Are user managers contacted for approval (policy exceptions)?",
        "Are incidents assigned severity based on data classification?",
        "Are incidents assigned to appropriate team (SOC, IR, Legal)?",
        "Is containment performed (block user, disable account, isolate endpoint)?",
        "Is investigation performed (correlate with other events)?",
        "Is evidence collected (logs, screenshots, file samples)?",
        "Is evidence preserved (chain of custody)?",
        "Are incidents reported to management (Critical/High)?",
        "Are incidents reported to DPO (personal data breaches)?",
        "Are incidents reported to CISO (all Critical incidents)?",
        "Are incidents reported to Legal (potential litigation)?",
        "Is remediation tracked (close incident when resolved)?",
        "Is Mean Time to Detect (MTTD) measured?",
        "Is Mean Time to Respond (MTTR) measured?",
        "Are lessons learned documented (post-incident review)?",
    ]

    # Row 5: sample row (first grey example — all cols F2F2F2)
    ws['A5'] = 1
    ws['B5'] = requirements[0]
    ws['C5'] = "Yes"
    ws['D5'] = "A812-4-IR-001"
    ws['E5'] = "ServiceNow integration complete"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}5'], styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    for row in range(6, 56):
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])

    # Apply data validation — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["response"].add(ws[f'C{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    row = 57  # secondary table starts at row 57 (gap row 56)

    # SLA Compliance Tracking Table
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "SLA COMPLIANCE TRACKING (Last 30 Days)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    sla_headers = ["Severity", "SLA Target", "Incidents (30d)", "Within SLA", "SLA Compliance %"]
    for col_idx, header in enumerate(sla_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    # Example rows
    sla_examples = [
        ("Critical", "15 minutes", 8, 7, "=D{}/C{}*100"),
        ("High", "1 hour", 35, 32, "=D{}/C{}*100"),
        ("Medium", "4 hours", 120, 115, "=D{}/C{}*100"),
        ("Low", "24 hours", 450, 448, "=D{}/C{}*100"),
    ]
    
    for severity, sla, incidents, within, formula in sla_examples:
        ws[f'A{row}'] = severity
        ws[f'B{row}'] = sla
        apply_style(ws[f'C{row}'], styles["input_cell"])
        apply_style(ws[f'D{row}'], styles["input_cell"])
        ws[f'C{row}'] = incidents
        ws[f'D{row}'] = within
        ws[f'E{row}'] = formula.format(row, row)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE

    # Freeze panes
    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 3
# ============================================================================

def create_soc_integration_sheet(ws, styles):
    """Create SOC Integration sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "SOC INTEGRATION ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess SOC alert workflow, escalation, and staffing (20 criteria)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Assessment criteria
    requirements = [
        "Are DLP alerts integrated with SOC workflow?",
        "Are Critical alerts delivered to SOC (SIEM + SMS/PagerDuty)?",
        "Are High alerts delivered to SOC (SIEM + Email)?",
        "Are SOC analysts trained on DLP alert triage?",
        "Is DLP training part of SOC analyst onboarding?",
        "Is DLP triage playbook documented (step-by-step)?",
        "Is escalation matrix documented (L1 → L2 → IR → Management)?",
        "Are SOC analysts able to escalate to DLP admin?",
        "Are SOC analysts able to escalate to user manager?",
        "Are SOC analysts able to escalate to Legal/DPO?",
        "Is SOC coverage 24/7 (for Critical DLP alerts)?",
        "Is SOC coverage business hours only (9-5) for Medium/Low alerts?",
        "Is on-call rotation defined (after-hours Critical alerts)?",
        "Are SOC metrics tracked (alert volume, triage time, SLA compliance)?",
        "Are SOC metrics reported to management (monthly)?",
        "Is SOC staffing adequate (not overwhelmed by alert volume)?",
        "Is SOC burnout monitored (high alert volume, turnover rate)?",
        "Are SOC tools adequate (SIEM, ticketing, communication)?",
        "Is SOC performance reviewed quarterly?",
        "Are SOC improvements implemented (based on performance review)?",
    ]

    # Row 5: sample row (first grey example — all cols F2F2F2)
    ws['A5'] = 1
    ws['B5'] = requirements[0]
    ws['C5'] = "Yes"
    ws['D5'] = "A812-4-SOC-001"
    ws['E5'] = "ServiceNow integration complete"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}5'], styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    for row in range(6, 56):
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])

    # Apply data validation — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["response"].add(ws[f'C{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE

    # Freeze panes
    ws.freeze_panes = 'A5'


def create_dashboards_reporting_sheet(ws, styles):
    """Create Dashboards Reporting sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "DASHBOARDS & REPORTING ASSESSMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess dashboard and reporting capabilities (20 criteria)"
    apply_style(ws['A2'], styles["subheader"])

    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Assessment criteria
    requirements = [
        "Is real-time SOC dashboard configured (alert queue)?",
        "SOC Dashboard: Shows Critical/High/Medium alerts pending triage?",
        "SOC Dashboard: Shows alert volume over time (hourly)?",
        "SOC Dashboard: Shows top blocked users (last 24 hours)?",
        "SOC Dashboard: Shows top blocked channels (email, USB, web)?",
        "SOC Dashboard: Shows false positive rate (rolling 7-day average)?",
        "Is executive dashboard configured (updated daily)?",
        "Executive Dashboard: Shows Critical incidents (last 30 days)?",
        "Executive Dashboard: Shows DLP effectiveness (prevented incidents)?",
        "Executive Dashboard: Shows compliance metrics (coverage %, enforcement rate)?",
        "Executive Dashboard: Shows risk heatmap (users, departments, channels)?",
        "Are weekly reports generated (Security Team)?",
        "Weekly Report: Alert summary (Critical/High/Medium/Low counts)?",
        "Weekly Report: Top users by alert volume?",
        "Weekly Report: False positive trend?",
        "Are monthly reports generated (Management)?",
        "Monthly Report: Executive summary (1-page, business language)?",
        "Monthly Report: Incident summary (critical events, resolved vs ongoing)?",
        "Are quarterly reports generated (CISO/Board)?",
        "Quarterly Report: Strategic risk assessment, ROI analysis?",
    ]

    # Row 5: sample row (first grey example — all cols F2F2F2)
    ws['A5'] = 1
    ws['B5'] = requirements[0]
    ws['C5'] = "Yes"
    ws['D5'] = "A812-4-DASH-001"
    ws['E5'] = "Splunk SOC dashboard live"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}5'], styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    for row in range(6, 56):
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])

    # Apply data validation — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["response"].add(ws[f'C{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE

    # Freeze panes
    ws.freeze_panes = 'A5'


def create_gap_analysis_sheet(ws, styles):
    """Create Gap Analysis sheet."""
    validations = create_data_validations()

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "GAP ANALYSIS - MONITORING & RESPONSE DEFICIENCIES"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35

    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Identify and track monitoring/response gaps and remediation plans (40 rows)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Row 3: empty spacing row (intentionally blank)

    # Column headers — row 4
    headers = [
        "Gap ID", "Gap Description", "Affected Area", "Risk Level",
        "Business Impact", "Remediation Plan", "Owner", "Target Date",
        "Status", "Evidence ID"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Row 5: sample row (first grey example — all cols F2F2F2)
    sample_gap = ("GAP-A812-4-001", "Critical alerts not reaching SOC SMS (delivery failures)", "Alerting",
                  "High", "Delayed response to data exfiltration", "Configure PagerDuty integration",
                  "SOC Lead", "2025-02-15", "In Progress", "A812-4-GAP-001")
    for col_idx, value in enumerate(sample_gap, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])

    # 50 empty FFFFCC rows starting at row 6
    for row in range(6, 56):
        for col_idx in range(1, 11):
            apply_style(ws.cell(row=row, column=col_idx), styles["input_cell"])

    # Apply data validations — rows 6-55 (after sample row 5)
    for r in range(6, 56):
        validations["affected_area"].add(ws[f'C{r}'])
        validations["risk"].add(ws[f'D{r}'])
        validations["gap_status"].add(ws[f'I{r}'])

    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['F'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['G'].width = WIDTH_WIDE
    ws.column_dimensions['H'].width = WIDTH_MEDIUM
    ws.column_dimensions['I'].width = WIDTH_MEDIUM
    ws.column_dimensions['J'].width = WIDTH_MEDIUM

    # Freeze panes
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A5'


def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — gold standard ER."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 — merge A1:H1, title "EVIDENCE REGISTER"
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for c in range(1, 9):
        ws.cell(row=1, column=c).border = border

    # Row 2 — subtitle, italic (NOT subheader/blue banner)
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers Row 4 — 8 columns, 003366 fill, bold white, thin borders
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Evidence Type dropdown (column C)
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Audit log,Compliance report,Dashboard export,Other"',
        allow_blank=True
    )
    ws.add_data_validation(ev_type_dv)

    # Verification Status dropdown (column H)
    ver_status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True
    )
    ws.add_data_validation(ver_status_dv)

    # F2F2F2 sample row at row 5
    INFO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = {
        1: "EV-001", 2: "Logging Configuration", 3: "Audit log",
        4: "SIEM log forwarding configuration — syslog rules export",
        5: "\\\\fileserver\\isms\\evidence\\dlp\\siem-log-config.pdf",
        6: "22.02.2026", 7: "IT Security Team", 8: "Verified"
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = INFO_FILL
        cell.border = border

    # Data rows 6-105 (100 rows)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Apply dropdowns to ranges
    ev_type_dv.add("C6:C105")
    ver_status_dv.add("H6:H105")

    # Column widths per spec
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 4
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard sheet — Gold Standard TABLE 1/2/3 layout."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title bar ────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle ─────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.12: Data Leakage Prevention"
    ws["A2"].font = Font(italic=True, size=10, color="003366", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: Blank spacer ─────────────────────────────────────────────────
    # (nothing written — intentionally blank)

    # ════════════════════════════════════════════════════════════════════════
    # TABLE 1 — COMPLIANCE ASSESSMENT SUMMARY
    # ════════════════════════════════════════════════════════════════════════

    # Row 4: TABLE 1 banner
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1 — COMPLIANCE ASSESSMENT SUMMARY"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 5: Column headers
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
                  "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = hdr
        cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # TABLE 1 data rows — one per assessment sheet
    # Sample row is at row 5 in each sheet; data starts at row 6.
    # Row 6:  Logging Configuration      — col C (response DV: Yes/No/Partial/Planned/N/A), rows 6-80 (25 req + 50 extra)
    # Row 7:  Alert Rules Inventory      — col L (rule_status DV: Active/Disabled/Testing/Deprecated), rows 6-59 (4 pre-pop + 50 empty)
    # Row 8:  SIEM Integration           — col C (response DV), rows 6-80 (25 req + 50 extra)
    # Row 9:  False Positive Management  — col C (response DV), rows 6-75 (20 req + 50 extra); FP rate table follows separately in lower rows
    # Row 10: Incident Response Workflow — col C (response DV), rows 6-80 (25 req + 50 extra)
    # Row 11: SOC Integration            — col C (response DV), rows 6-75 (20 req + 50 extra)
    # Row 12: Dashboards & Reporting     — col C (response DV), rows 6-75 (20 req + 50 extra)
    # Row 13: Gap Analysis               — col I (gap_status DV: Not Started/In Progress/Complete/Blocked), rows 6-57 (2 pre-pop + 50 empty)

    t1_data = [
        # (label, total_formula, compliant_formula, partial_formula, nc_formula, na_formula)
        (
            "Logging Configuration",
            "=COUNTA('Logging Configuration'!C6:C55)-COUNTIF('Logging Configuration'!C6:C55,\"\")",
            "=COUNTIF('Logging Configuration'!C6:C55,\"Yes\")",
            "=COUNTIF('Logging Configuration'!C6:C55,\"Partial\")+COUNTIF('Logging Configuration'!C6:C55,\"Planned\")",
            "=COUNTIF('Logging Configuration'!C6:C55,\"No\")",
            "=COUNTIF('Logging Configuration'!C6:C55,\"N/A\")",
        ),
        (
            "Alert Rules Inventory",
            "=COUNTA('Alert Rules Inventory'!L6:L55)-COUNTIF('Alert Rules Inventory'!L6:L55,\"\")",
            "=COUNTIF('Alert Rules Inventory'!L6:L55,\"Active\")",
            "=COUNTIF('Alert Rules Inventory'!L6:L55,\"Testing\")",
            "=COUNTIF('Alert Rules Inventory'!L6:L55,\"Disabled\")+COUNTIF('Alert Rules Inventory'!L6:L55,\"Deprecated\")",
            "0",
        ),
        (
            "SIEM Integration",
            "=COUNTA('SIEM Integration'!C6:C55)-COUNTIF('SIEM Integration'!C6:C55,\"\")",
            "=COUNTIF('SIEM Integration'!C6:C55,\"Yes\")",
            "=COUNTIF('SIEM Integration'!C6:C55,\"Partial\")+COUNTIF('SIEM Integration'!C6:C55,\"Planned\")",
            "=COUNTIF('SIEM Integration'!C6:C55,\"No\")",
            "=COUNTIF('SIEM Integration'!C6:C55,\"N/A\")",
        ),
        (
            "False Positive Management",
            "=COUNTA('False Positive Management'!C6:C55)-COUNTIF('False Positive Management'!C6:C55,\"\")",
            "=COUNTIF('False Positive Management'!C6:C55,\"Yes\")",
            "=COUNTIF('False Positive Management'!C6:C55,\"Partial\")+COUNTIF('False Positive Management'!C6:C55,\"Planned\")",
            "=COUNTIF('False Positive Management'!C6:C55,\"No\")",
            "=COUNTIF('False Positive Management'!C6:C55,\"N/A\")",
        ),
        (
            "Incident Response Workflow",
            "=COUNTA('Incident Response Workflow'!C6:C55)-COUNTIF('Incident Response Workflow'!C6:C55,\"\")",
            "=COUNTIF('Incident Response Workflow'!C6:C55,\"Yes\")",
            "=COUNTIF('Incident Response Workflow'!C6:C55,\"Partial\")+COUNTIF('Incident Response Workflow'!C6:C55,\"Planned\")",
            "=COUNTIF('Incident Response Workflow'!C6:C55,\"No\")",
            "=COUNTIF('Incident Response Workflow'!C6:C55,\"N/A\")",
        ),
        (
            "SOC Integration",
            "=COUNTA('SOC Integration'!C6:C55)-COUNTIF('SOC Integration'!C6:C55,\"\")",
            "=COUNTIF('SOC Integration'!C6:C55,\"Yes\")",
            "=COUNTIF('SOC Integration'!C6:C55,\"Partial\")+COUNTIF('SOC Integration'!C6:C55,\"Planned\")",
            "=COUNTIF('SOC Integration'!C6:C55,\"No\")",
            "=COUNTIF('SOC Integration'!C6:C55,\"N/A\")",
        ),
        (
            "Dashboards & Reporting",
            "=COUNTA('Dashboards Reporting'!C6:C55)-COUNTIF('Dashboards Reporting'!C6:C55,\"\")",
            "=COUNTIF('Dashboards Reporting'!C6:C55,\"Yes\")",
            "=COUNTIF('Dashboards Reporting'!C6:C55,\"Partial\")+COUNTIF('Dashboards Reporting'!C6:C55,\"Planned\")",
            "=COUNTIF('Dashboards Reporting'!C6:C55,\"No\")",
            "=COUNTIF('Dashboards Reporting'!C6:C55,\"N/A\")",
        ),
        (
            "Gap Analysis",
            "=COUNTA('Gap Analysis'!I6:I55)-COUNTIF('Gap Analysis'!I6:I55,\"\")",
            "=COUNTIF('Gap Analysis'!I6:I55,\"Complete\")",
            "=COUNTIF('Gap Analysis'!I6:I55,\"In Progress\")",
            "=COUNTIF('Gap Analysis'!I6:I55,\"Not Started\")+COUNTIF('Gap Analysis'!I6:I55,\"Blocked\")",
            "0",
        ),
    ]

    data_start = 6  # first data row
    for i, (label, total_f, comp_f, part_f, nc_f, na_f) in enumerate(t1_data):
        r = data_start + i
        # Col A: label
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        # Col B: Total Items = SUM(C+D+E+F) — computed from compliant+partial+nc+na
        ws.cell(row=r, column=2).value = f"=C{r}+D{r}+E{r}+F{r}"
        ws.cell(row=r, column=2).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).border = border
        # Col C: Compliant
        ws.cell(row=r, column=3).value = comp_f
        ws.cell(row=r, column=3).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).border = border
        # Col D: Partial
        ws.cell(row=r, column=4).value = part_f
        ws.cell(row=r, column=4).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4).border = border
        # Col E: Non-Compliant
        ws.cell(row=r, column=5).value = nc_f
        ws.cell(row=r, column=5).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=5).border = border
        # Col F: N/A
        ws.cell(row=r, column=6).value = na_f
        ws.cell(row=r, column=6).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=6).border = border
        # Col G: Compliance % = IF((B-F)=0,0,C/(B-F))
        ws.cell(row=r, column=7).value = f"=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))"
        ws.cell(row=r, column=7).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=7).border = border

    # TOTAL row (row 14 = data_start + 8)
    total_row = data_start + len(t1_data)  # = 14
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    for col_idx in range(2, 7):
        col_letter = ["B", "C", "D", "E", "F"][col_idx - 2]
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})"
        cell.font = Font(bold=True, size=10, name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # G14: Overall compliance % (used by Approval Sign-Off)
    ws.cell(row=total_row, column=7).value = (
        f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    )
    ws.cell(row=total_row, column=7).font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=total_row, column=7).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=7).number_format = "0.0%"
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=7).border = border

    # ════════════════════════════════════════════════════════════════════════
    # TABLE 2 — KEY METRICS
    # ════════════════════════════════════════════════════════════════════════

    t2_start = total_row + 2  # row 16

    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2 — KEY METRICS"
    ws[f"A{t2_start}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{t2_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{t2_start}"].alignment = Alignment(horizontal="left", vertical="center")

    # TABLE 2 headers
    t2_hdr_row = t2_start + 1  # row 17
    t2_headers = ["Metric", "Value", "Notes"]
    for col_idx, hdr in enumerate(t2_headers, start=1):
        cell = ws.cell(row=t2_hdr_row, column=col_idx)
        cell.value = hdr
        cell.font = Font(bold=True, size=10, name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Merge C across remaining cols for "Notes" header visual
    ws.merge_cells(f"C{t2_hdr_row}:G{t2_hdr_row}")

    # TABLE 2 data
    t2_metrics = [
        ("Overall Compliance Score",
         f"=G{total_row}",
         "Compliant items / (Total - N/A items)"),
        ("Critical Gaps (Risk = High/Critical)",
         "=COUNTIF('Gap Analysis'!D6:D55,\"High\")+COUNTIF('Gap Analysis'!D6:D55,\"Critical\")",
         "Open high-risk monitoring/response gaps"),
        ("Logging Coverage (all channels)",
         f"=IF((B{data_start}-F{data_start})=0,0,C{data_start}/(B{data_start}-F{data_start}))",
         "Logging Configuration compliance rate"),
        ("SIEM Integration Coverage",
         f"=IF((B{data_start+2}-F{data_start+2})=0,0,C{data_start+2}/(B{data_start+2}-F{data_start+2}))",
         "SIEM Integration compliance rate"),
        ("Incident Response Readiness",
         f"=IF((B{data_start+4}-F{data_start+4})=0,0,C{data_start+4}/(B{data_start+4}-F{data_start+4}))",
         "IR Workflow compliance rate"),
        ("Gaps Resolved",
         "=COUNTIF('Gap Analysis'!I6:I55,\"Complete\")",
         "Monitoring/response gaps fully closed"),
    ]

    t2_data_start = t2_hdr_row + 1  # row 18
    for j, (metric, formula, note) in enumerate(t2_metrics):
        r = t2_data_start + j
        ws.cell(row=r, column=1).value = metric
        ws.cell(row=r, column=1).font = Font(size=10, name="Calibri")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(row=r, column=2).value = formula
        ws.cell(row=r, column=2).font = Font(size=10, color="000000", name="Calibri")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        # Format as % for rows that are compliance rates, plain for counts
        if "Score" in metric or "Coverage" in metric or "Readiness" in metric:
            ws.cell(row=r, column=2).number_format = "0.0%"

        ws.merge_cells(f"C{r}:G{r}")
        ws.cell(row=r, column=3).value = note
        ws.cell(row=r, column=3).font = Font(size=10, name="Calibri")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center")

    # ════════════════════════════════════════════════════════════════════════
    # TABLE 3 — KEY FINDINGS
    # ════════════════════════════════════════════════════════════════════════

    t3_start = t2_data_start + len(t2_metrics) + 2  # 2-row gap

    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3 — KEY FINDINGS"
    ws[f"A{t3_start}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{t3_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{t3_start}"].alignment = Alignment(horizontal="left", vertical="center")

    t3_hdr_row = t3_start + 1
    t3_hdrs = ["#", "Finding"]
    for col_idx, hdr in enumerate(t3_hdrs, start=1):
        cell = ws.cell(row=t3_hdr_row, column=col_idx)
        cell.value = hdr
        cell.font = Font(bold=True, size=10, name="Calibri")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.merge_cells(f"B{t3_hdr_row}:G{t3_hdr_row}")

    # 5 FFFFCC finding rows
    t3_data_start = t3_hdr_row + 1
    for k in range(1, 6):
        r = t3_data_start + k - 1
        num_cell = ws.cell(row=r, column=1)
        num_cell.value = k
        num_cell.font = Font(size=10, name="Calibri")
        num_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        num_cell.alignment = Alignment(horizontal="center", vertical="center")
        num_cell.border = border

        ws.merge_cells(f"B{r}:G{r}")
        finding_cell = ws.cell(row=r, column=2)
        finding_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        finding_cell.border = border
        finding_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 5 (APPROVAL SIGN-OFF)
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet — gold standard AS."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 — merge A1:E1, "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = f"ISO/IEC 27001:2022 \u2014 Control A.8.12: {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner — 4472C4 fill
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G13),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Status dropdown on Assessment Status
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)
    if status_row:
        status_dv.add(ws[f"B{status_row}"])

    # 3 approver sections (gap of 2 rows before first)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True, name="Calibri")
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION row — immediately after last approver
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS — gap of 3 rows
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 5: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the workbook."""
    logger.info(f"Generating {WORKBOOK_ID} - {ASSESSMENT_AREA}...")

    # Create workbook and styles
    wb = create_workbook()
    styles = _STYLES

    # Create all sheets (each function creates its own validations)
    logger.info("  [1/13] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("  [2/13] Creating Logging Configuration...")
    create_logging_configuration_sheet(wb["Logging Configuration"], styles)

    logger.info("  [3/13] Creating Alert Rules Inventory...")
    create_alert_rules_inventory_sheet(wb["Alert Rules Inventory"], styles)

    logger.info("  [4/13] Creating Alert Volume Metrics...")
    create_alert_volume_metrics_sheet(wb["Alert Volume Metrics"], styles)

    logger.info("  [5/13] Creating SIEM Integration...")
    create_siem_integration_sheet(wb["SIEM Integration"], styles)

    logger.info("  [6/13] Creating False Positive Management...")
    create_false_positive_management_sheet(wb["False Positive Management"], styles)

    logger.info("  [7/13] Creating Incident Response Workflow...")
    create_incident_response_workflow_sheet(wb["Incident Response Workflow"], styles)

    logger.info("  [8/13] Creating SOC Integration...")
    create_soc_integration_sheet(wb["SOC Integration"], styles)

    logger.info("  [9/13] Creating Dashboards Reporting...")
    create_dashboards_reporting_sheet(wb["Dashboards Reporting"], styles)

    logger.info("  [10/13] Creating Gap Analysis...")
    create_gap_analysis_sheet(wb["Gap Analysis"], styles)

    logger.info("  [11/13] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("  [12/13] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("  [13/13] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)

    # Save workbook
    filename = f"ISMS-IMP-A.8.12.4_Monitoring_Response_{datetime.now().strftime('%Y%m%d')}.xlsx"
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"\u2705 Success! Workbook saved as: {filename}")
    logger.info(f"Workbook Statistics:")
    logger.info(f"  - Total Sheets: 13")
    logger.info(f"  - Assessment Items: ~135 monitoring/response criteria")
    logger.info(f"  - Alert Rules: 30 rows")
    logger.info(f"  - Gap Analysis: 40 rows")
    logger.info(f"  - Evidence Register: 100 rows")
    logger.info(f"  - Approval Sign-Off: 3-level workflow")


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
