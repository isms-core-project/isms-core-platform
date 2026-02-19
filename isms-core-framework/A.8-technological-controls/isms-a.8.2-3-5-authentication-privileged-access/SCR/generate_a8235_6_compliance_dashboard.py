#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.2-3-5.6 - Authentication & Privileged Access Compliance Dashboard
================================================================================

ISO/IEC 27001:2022 Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access
Consolidation Dashboard: Executive Compliance Overview

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dashboard requirements, KPI definitions, and
reporting structure.

Key customization areas:
1. External workbook paths (update to your actual file locations)
2. Compliance scoring weights (per your governance requirements)
3. KPI thresholds (MFA coverage, PAM vault per your targets)
4. Regulatory mappings (NIS2, DORA per your jurisdictions)
5. Approval workflow (per your organizational structure)

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Consolidates data from all five authentication and privileged access assessment
domains into a unified executive dashboard, providing comprehensive compliance
visibility for ISO 27001:2022 Controls A.8.2, A.8.3, and A.8.5.

**Consolidated Data Sources:**
- Domain 1: Authentication Inventory Assessment (ISMS-IMP-A.8.2-3-5.S1)
- Domain 2: MFA Coverage Assessment ()
- Domain 3: Privileged Account Inventory Assessment (ISMS-IMP-A.8.2-3-5.S3)
- Domain 4: Privileged Access Monitoring Assessment (ISMS-IMP-A.8.2-3-5.S4)
- Domain 5: Information Access Restriction Assessment (ISMS-IMP-A.8.2-3-5.S5)

**Generated Dashboard Structure:**
1. Executive Dashboard - Overall compliance status, KPIs, critical gaps
2. Gap Analysis - Consolidated gaps from all 5 domains (200 rows)
3. Risk Register - Authentication & PAM risks (100 rows)
4. Remediation Roadmap - Action plans with owners/timeline (200 rows)
5. KPIs & Metrics - Detailed KPI tracking
6. Evidence Register - Audit evidence index (100 rows)
7. Action Items & Follow-up - Open action tracking (100 rows)
8. Audit & Compliance Log - Audit trail
9. Approval Sign-Off - Executive approval workflow

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



# ============================================================================
# CONSTANTS
# ============================================================================

CHECK = '\u2705'      # ✅
XMARK = '\u274C'      # ❌
WARNING = '\u26A0'    # ⚠️
BULLET = '\u2022'     # •

DOCUMENT_ID = "ISMS-IMP-A.8.2-3-5.6"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Compliance_Dashboard_{GENERATED_TIMESTAMP}.xlsx"
CONTROL_REF = "ISO/IEC 27001:2022 A.8.2, A.8.3, A.8.5"

# Source workbook names (normalized)
SOURCE_WORKBOOKS = {
    "auth_inventory": "ISMS-IMP-A.8.2-3-5.1.xlsx",
    "mfa_coverage": "ISMS-IMP-A.8.2-3-5.2.xlsx",
    "privileged_accounts": "ISMS-IMP-A.8.2-3-5.3.xlsx",
    "access_monitoring": "ISMS-IMP-A.8.2-3-5.4.xlsx",
    "access_restriction": "ISMS-IMP-A.8.2-3-5.5.xlsx",
}


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook():
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Executive Dashboard",
        "Gap Analysis",
        "Risk Register",
        "Remediation Roadmap",
        "KPIs & Metrics",
        "Evidence Register",
        "Action Items & Follow-up",
        "Audit & Compliance Log",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "critical_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_yellow": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: EXECUTIVE DASHBOARD
# ============================================================================

def create_executive_dashboard(ws, styles):
    """Create Executive Dashboard with external workbook links and KPIs."""

    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "ISMS-IMP-A.8.2-3-5.6 – Authentication & Privileged Access Compliance Dashboard\n"
        "ISO/IEC 27001:2022 - Controls A.8.2, A.8.3, A.8.5 - Executive Overview"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # ---------- DOCUMENT INFORMATION ----------
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Report Type", "Compliance Summary Dashboard"),
        ("Related Policy", "ISMS-POL-A.8.2-3-5 (Authentication & Access)"),
        ("Version", "1.0"),
        ("Report Date", ""),
        ("Reporting Period", ""),
        ("Prepared By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
        ("Last Updated", "=TODAY()"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        elif "=" in str(value):
            ws[f"B{row}"].font = Font(color="0000FF")
        row += 1

    # ---------- OVERALL COMPLIANCE STATUS ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "OVERALL AUTHENTICATION & ACCESS COMPLIANCE STATUS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    # Compliance Scorecard
    row += 1
    scorecard_headers = ["Metric", "Score", "Target", "Status"]
    for col_idx, header in enumerate(scorecard_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    scorecard_metrics = [
        ("Overall Compliance Rate", '=SUMPRODUCT(VALUE(SUBSTITUTE(C26:C30,"%","")))/COUNTA(C26:C30)&"%"', "≥95%", ""),
        ("Critical Gaps", "=COUNTIFS('Gap Analysis'!I8:I207,\"Critical\",'Gap Analysis'!R8:R207,\"<>Closed\")", "0", ""),
        ("High-Risk Items", "=COUNTIFS('Risk Register'!N21:N120,\">=15\",'Risk Register'!Q21:Q120,\"<>Closed\")", "≤5", ""),
        ("Remediation Progress", "=COUNTIF('Remediation Roadmap'!H37:H236,\"Completed\")/COUNTA('Remediation Roadmap'!H37:H236)*100&\"%\"", "≥80%", ""),
    ]

    row += 1
    for metric, formula, target, status in scorecard_metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        ws.cell(row=row, column=2, value=formula).font = Font(bold=True, color="0000FF")
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    # Add dropdown for status column
    dv_status = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(f"D{row-4}:D{row-1}")

    # ---------- COMPLIANCE BY ASSESSMENT AREA ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "COMPLIANCE BY ASSESSMENT AREA"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    area_headers = [
        "Assessment Area",
        "Source Document",
        "Overall Compliance %",
        "Total Items",
        "Compliant",
        "Partial",
        "Non-Compliant",
        "Trend"
    ]

    for col_idx, header in enumerate(area_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Column widths
    widths = [28, 24, 18, 12, 12, 12, 14, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows with EXTERNAL WORKBOOK LINKS
    assessment_areas = [
        ("A.8.5 - Authentication Inventory", SOURCE_WORKBOOKS["auth_inventory"], 16),
        ("A.8.5 - MFA Coverage", SOURCE_WORKBOOKS["mfa_coverage"], 11),
        ("A.8.2 - Privileged Accounts", SOURCE_WORKBOOKS["privileged_accounts"], 11),
        ("A.8.2 - Access Monitoring", SOURCE_WORKBOOKS["access_monitoring"], 9),
        ("A.8.3 - Access Restriction", SOURCE_WORKBOOKS["access_restriction"], 9),
    ]

    row += 1
    area_start_row = row
    for area_name, source_doc, total_row in assessment_areas:
        ws.cell(row=row, column=1, value=area_name)
        ws.cell(row=row, column=2, value=source_doc)

        # EXTERNAL LINKS - formulas reference normalized source workbooks
        ws.cell(row=row, column=3, value=f"='[{source_doc}]Summary Dashboard'!G{total_row}")
        ws.cell(row=row, column=4, value=f"='[{source_doc}]Summary Dashboard'!B{total_row}")
        ws.cell(row=row, column=5, value=f"='[{source_doc}]Summary Dashboard'!C{total_row}")
        ws.cell(row=row, column=6, value=f"='[{source_doc}]Summary Dashboard'!D{total_row}")
        ws.cell(row=row, column=7, value=f"='[{source_doc}]Summary Dashboard'!E{total_row}")

        # Trend - manual entry with dropdown
        ws.cell(row=row, column=8).fill = styles["input_cell"]["fill"]

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles["border"]

        row += 1

    # Trend dropdown
    dv_trend = DataValidation(type="list", formula1='"↑ Improved,→ No Change,↓ Declined"', allow_blank=False)
    ws.add_data_validation(dv_trend)
    for r in range(area_start_row, row):
        dv_trend.add(ws.cell(row=r, column=8))

    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value="(Aggregated)").font = Font(italic=True)
    ws.cell(row=row, column=3, value=f'=SUMPRODUCT(VALUE(SUBSTITUTE(C{area_start_row}:C{row-1},"%","")))/COUNTA(C{area_start_row}:C{row-1})&"%"').font = Font(bold=True, color="0000FF", size=12)
    ws.cell(row=row, column=4, value=f"=SUM(D{area_start_row}:D{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E{area_start_row}:E{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{area_start_row}:F{row-1})").font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"=SUM(G{area_start_row}:G{row-1})").font = Font(bold=True)

    for col in range(1, 8):
        ws.cell(row=row, column=col).border = styles["border"]

    # ---------- KEY PERFORMANCE INDICATORS ----------
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "KEY PERFORMANCE INDICATORS (KPIs)"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    kpi_headers = ["KPI", "Current Value", "Target", "Status", "Last Quarter", "Change"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    kpi_metrics = [
        ("MFA Coverage (All Users)", "", "≥85%", "", "", ""),
        ("MFA Coverage (Privileged)", "", "100%", "", "", ""),
        ("MFA Coverage (Remote Access)", "", "100%", "", "", ""),
        ("SSO Adoption Rate", "", "≥80%", "", "", ""),
        ("Accounts in PAM Vault", "", "100%", "", "", ""),
        ("Session Recording (Tier 0)", "", "100%", "", "", ""),
        ("Tier Isolation Violations", "", "0", "", "", ""),
        ("Access Review Completion", "", "100%", "", "", ""),
        ("Orphaned Account Count", "", "0", "", "", ""),
        ("Open Critical Gaps", f"=COUNTIFS('Gap Analysis'!I8:I207,\"Critical\",'Gap Analysis'!R8:R207,\"<>Closed\")", "0", "", "", ""),
    ]

    row += 1
    for kpi_name, current, target, status, last_q, change in kpi_metrics:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=2, value=current)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]

        if "=" in str(current):
            ws.cell(row=row, column=2).font = Font(color="0000FF")
        else:
            ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = styles["border"]

        row += 1

    # ---------- TOP 5 CRITICAL ISSUES ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "TOP 5 CRITICAL SECURITY GAPS"
    ws[f"A{row}"].font = styles["critical_header"]["font"]
    ws[f"A{row}"].fill = styles["critical_header"]["fill"]
    ws[f"A{row}"].alignment = styles["critical_header"]["alignment"]

    row += 1
    critical_headers = ["Rank", "Issue Description", "Control Area", "Risk Level", "Systems Affected", "Target Date", "Owner", "Status"]
    for col_idx, header in enumerate(critical_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for rank in range(1, 6):
        ws.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = styles["border"]
        for col in range(2, 9):
            ws.cell(row=row, column=col).fill = styles["input_cell"]["fill"]
            ws.cell(row=row, column=col).border = styles["border"]
        row += 1

    # ---------- EXECUTIVE SUMMARY ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "EXECUTIVE SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    summary_fields = [
        ("Assessment Period:", ""),
        ("Overall Compliance Status:", ""),
        ("Security Posture:", ""),
    ]

    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "Key Achievements This Period:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws.merge_cells(f"A{row}:H{row+4}")
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{row}"].border = styles["border"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row += 6
    ws[f"A{row}"] = "Major Concerns:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws.merge_cells(f"A{row}:H{row+4}")
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{row}"].border = styles["border"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row += 6
    ws[f"A{row}"] = "Recommended Executive Actions:"
    ws[f"A{row}"].font = Font(bold=True)

    for i in range(1, 4):
        row += 1
        ws[f"A{row}"] = f"{i}."
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:H{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 3: GAP ANALYSIS
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create comprehensive gap analysis sheet with 200 data rows."""

    # Header
    ws.merge_cells("A1:T1")
    ws["A1"] = (
        "COMPREHENSIVE GAP ANALYSIS - AUTHENTICATION & ACCESS CONTROLS\n"
        "Consolidated view of all compliance gaps across assessment areas"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Summary Statistics Section
    ws["A3"] = "GAP SUMMARY STATISTICS"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    summary_stats = [
        ("Total Gaps Identified:", "=COUNTA(A8:A207)-COUNTBLANK(A8:A207)"),
        ("Critical Gaps:", "=COUNTIF(I8:I207,\"Critical\")"),
        ("High Gaps:", "=COUNTIF(I8:I207,\"High\")"),
        ("Medium Gaps:", "=COUNTIF(I8:I207,\"Medium\")"),
        ("Low Gaps:", "=COUNTIF(I8:I207,\"Low\")"),
        ("Closed Gaps:", "=COUNTIF(R8:R207,\"Closed\")"),
        ("Open Gaps:", "=COUNTA(A8:A207)-COUNTBLANK(A8:A207)-COUNTIF(R8:R207,\"Closed\")"),
    ]

    for label, formula in summary_stats:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    # Column Headers
    row = 7
    gap_headers = [
        ("A", "Gap ID", 12),
        ("B", "Assessment Area", 22),
        ("C", "Control Reference", 18),
        ("D", "Gap Description", 45),
        ("E", "Expected State", 30),
        ("F", "Current State", 30),
        ("G", "Root Cause", 25),
        ("H", "Impact", 25),
        ("I", "Severity", 12),
        ("J", "Likelihood", 12),
        ("K", "Risk Score", 12),
        ("L", "Affected Systems", 20),
        ("M", "Recommended Action", 35),
        ("N", "Owner", 18),
        ("O", "Target Date", 14),
        ("P", "Estimated Cost", 14),
        ("Q", "Evidence Ref", 15),
        ("R", "Status", 12),
        ("S", "Last Updated", 14),
        ("T", "Comments", 30),
    ]

    for col_letter, header, width in gap_headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[col_letter].width = width

    # Data rows with validations
    dv_severity = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    dv_likelihood = DataValidation(type="list", formula1='"Very High,High,Medium,Low,Very Low"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Open,In Progress,Pending Approval,Closed,Deferred"', allow_blank=True)
    dv_area = DataValidation(type="list", formula1='"A.8.5 Authentication,A.8.5 MFA,A.8.2 Privileged Access,A.8.2 Monitoring,A.8.3 Access Restriction"', allow_blank=True)

    ws.add_data_validation(dv_severity)
    ws.add_data_validation(dv_likelihood)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_area)

    # Pre-populate 200 data rows
    for i in range(200):
        data_row = row + 1 + i
        ws[f"A{data_row}"] = f"GAP-{i+1:03d}" if i < 10 else ""  # Pre-fill first 10 IDs
        ws[f"A{data_row}"].font = Font(color="808080") if i >= 10 else Font()

        # Risk score formula
        ws[f"K{data_row}"] = f'=IF(OR(I{data_row}="",J{data_row}=""),"",VLOOKUP(I{data_row},{{\"Critical\",5;\"High\",4;\"Medium\",3;\"Low\",2}},2,FALSE)*VLOOKUP(J{data_row},{{\"Very High\",5;\"High\",4;\"Medium\",3;\"Low\",2;\"Very Low\",1}},2,FALSE))'

        # Apply input cell formatting
        for col in ["B", "C", "D", "E", "F", "G", "H", "L", "M", "N", "O", "P", "Q", "S", "T"]:
            ws[f"{col}{data_row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{data_row}"].border = styles["border"]

        # Apply validations
        dv_area.add(ws[f"B{data_row}"])
        dv_severity.add(ws[f"I{data_row}"])
        dv_likelihood.add(ws[f"J{data_row}"])
        dv_status.add(ws[f"R{data_row}"])

        # Style specific columns
        ws[f"I{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{data_row}"].border = styles["border"]
        ws[f"J{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{data_row}"].border = styles["border"]
        ws[f"R{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"R{data_row}"].border = styles["border"]

    ws.freeze_panes = "A8"


# ============================================================================
# SECTION 4: RISK REGISTER
# ============================================================================

def create_risk_register(ws, styles):
    """Create risk register with 100 data rows."""

    # Header
    ws.merge_cells("A1:Q1")
    ws["A1"] = (
        "AUTHENTICATION & ACCESS RISK REGISTER\n"
        "Residual risks requiring ongoing management and monitoring"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Risk Matrix Summary
    ws["A3"] = "RISK MATRIX SUMMARY"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    risk_stats = [
        ("Total Risks:", "=COUNTA(A21:A120)-COUNTBLANK(A21:A120)"),
        ("Critical Risks (≥20):", "=COUNTIF(N21:N120,\">=20\")"),
        ("High Risks (15-19):", "=COUNTIFS(N21:N120,\">=15\",N21:N120,\"<20\")"),
        ("Medium Risks (10-14):", "=COUNTIFS(N21:N120,\">=10\",N21:N120,\"<15\")"),
        ("Low Risks (<10):", "=COUNTIF(N21:N120,\"<10\")"),
        ("Open Risks:", "=COUNTA(A21:A120)-COUNTBLANK(A21:A120)-COUNTIF(Q21:Q120,\"Closed\")"),
    ]

    for label, formula in risk_stats:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    # Risk Matrix Visual (simplified)
    ws["D4"] = "RISK MATRIX"
    ws["D4"].font = Font(bold=True)
    ws.merge_cells("D5:H9")
    ws["D5"] = "Impact vs Likelihood Matrix\n[Visual representation - see standard risk matrix]"
    ws["D5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D5"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Column Headers
    row = 20
    risk_headers = [
        ("A", "Risk ID", 12),
        ("B", "Risk Category", 18),
        ("C", "Control Area", 18),
        ("D", "Risk Description", 40),
        ("E", "Risk Owner", 18),
        ("F", "Threat Source", 20),
        ("G", "Vulnerability", 25),
        ("H", "Existing Controls", 30),
        ("I", "Control Effectiveness", 18),
        ("J", "Impact (1-5)", 12),
        ("K", "Likelihood (1-5)", 14),
        ("L", "Inherent Risk", 14),
        ("M", "Mitigating Actions", 30),
        ("N", "Residual Risk", 14),
        ("O", "Risk Treatment", 16),
        ("P", "Target Date", 14),
        ("Q", "Status", 12),
    ]

    for col_letter, header, width in risk_headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[col_letter].width = width

    # Validations
    dv_category = DataValidation(type="list", formula1='"Identity,Authentication,Authorization,Privileged Access,Monitoring,Compliance"', allow_blank=True)
    dv_effectiveness = DataValidation(type="list", formula1='"Fully Effective,Partially Effective,Not Effective,Not Implemented"', allow_blank=True)
    dv_treatment = DataValidation(type="list", formula1='"Accept,Mitigate,Transfer,Avoid"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Open,In Treatment,Monitoring,Closed"', allow_blank=True)
    dv_impact = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_likelihood = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)

    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_effectiveness)
    ws.add_data_validation(dv_treatment)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_impact)
    ws.add_data_validation(dv_likelihood)

    # Pre-populate 100 data rows
    for i in range(100):
        data_row = row + 1 + i
        ws[f"A{data_row}"] = f"RISK-{i+1:03d}" if i < 5 else ""
        ws[f"A{data_row}"].font = Font(color="808080") if i >= 5 else Font()

        # Inherent risk formula
        ws[f"L{data_row}"] = f"=IF(OR(J{data_row}=\"\",K{data_row}=\"\"),\"\",J{data_row}*K{data_row})"

        # Apply input formatting
        for col in ["B", "C", "D", "E", "F", "G", "H", "M", "N", "O", "P"]:
            ws[f"{col}{data_row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{data_row}"].border = styles["border"]

        # Apply validations
        dv_category.add(ws[f"B{data_row}"])
        dv_effectiveness.add(ws[f"I{data_row}"])
        dv_impact.add(ws[f"J{data_row}"])
        dv_likelihood.add(ws[f"K{data_row}"])
        dv_treatment.add(ws[f"O{data_row}"])
        dv_status.add(ws[f"Q{data_row}"])

        ws[f"I{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{data_row}"].border = styles["border"]
        ws[f"J{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"J{data_row}"].border = styles["border"]
        ws[f"K{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{data_row}"].border = styles["border"]
        ws[f"Q{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"Q{data_row}"].border = styles["border"]

    ws.freeze_panes = "A21"


# ============================================================================
# SECTION 5: REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap(ws, styles):
    """Create remediation roadmap with 200 data rows."""

    # Header
    ws.merge_cells("A1:R1")
    ws["A1"] = (
        "REMEDIATION ROADMAP - AUTHENTICATION & ACCESS CONTROLS\n"
        "Prioritised action plan for achieving compliance"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Progress Summary
    ws["A3"] = "REMEDIATION PROGRESS"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    progress_stats = [
        ("Total Remediation Items:", "=COUNTA(A37:A236)-COUNTBLANK(A37:A236)"),
        ("Completed:", "=COUNTIF(H37:H236,\"Completed\")"),
        ("In Progress:", "=COUNTIF(H37:H236,\"In Progress\")"),
        ("Not Started:", "=COUNTIF(H37:H236,\"Not Started\")"),
        ("Blocked:", "=COUNTIF(H37:H236,\"Blocked\")"),
        ("Completion Rate:", "=IF(COUNTA(A37:A236)-COUNTBLANK(A37:A236)=0,\"0%\",ROUND(COUNTIF(H37:H236,\"Completed\")/(COUNTA(A37:A236)-COUNTBLANK(A37:A236))*100,1)&\"%\")"),
    ]

    for label, formula in progress_stats:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    # Timeline Summary
    ws["D4"] = "BY PRIORITY"
    ws["D4"].font = Font(bold=True)
    ws["D5"] = "Critical:"
    ws["E5"] = "=COUNTIF(D37:D236,\"Critical\")"
    ws["E5"].font = Font(color="0000FF")
    ws["D6"] = "High:"
    ws["E6"] = "=COUNTIF(D37:D236,\"High\")"
    ws["E6"].font = Font(color="0000FF")
    ws["D7"] = "Medium:"
    ws["E7"] = "=COUNTIF(D37:D236,\"Medium\")"
    ws["E7"].font = Font(color="0000FF")
    ws["D8"] = "Low:"
    ws["E8"] = "=COUNTIF(D37:D236,\"Low\")"
    ws["E8"].font = Font(color="0000FF")

    # Budget Summary
    ws["G4"] = "BUDGET SUMMARY"
    ws["G4"].font = Font(bold=True)
    ws["G5"] = "Total Estimated:"
    ws["H5"] = "=SUM(R37:R236)"
    ws["H5"].font = Font(bold=True, color="0000FF")
    ws["H5"].number_format = '"CHF"#,##0'
    ws["G6"] = "Approved Budget:"
    ws["H6"].fill = styles["input_cell"]["fill"]
    ws["H6"].number_format = '"CHF"#,##0'
    ws["G7"] = "Budget Gap:"
    ws["H7"] = "=H6-H5"
    ws["H7"].font = Font(color="0000FF")
    ws["H7"].number_format = '"CHF"#,##0'

    # Column Headers
    row = 36
    roadmap_headers = [
        ("A", "Item ID", 12),
        ("B", "Gap Reference", 14),
        ("C", "Description", 40),
        ("D", "Priority", 12),
        ("E", "Control Area", 18),
        ("F", "Owner", 18),
        ("G", "Accountable", 18),
        ("H", "Status", 14),
        ("I", "Start Date", 14),
        ("J", "Target Date", 14),
        ("K", "Actual Completion", 16),
        ("L", "% Complete", 12),
        ("M", "Dependencies", 20),
        ("N", "Milestones", 25),
        ("O", "Resources Required", 20),
        ("P", "Risk if Delayed", 20),
        ("Q", "Evidence Ref", 15),
        ("R", "Estimated Cost", 14),
    ]

    for col_letter, header, width in roadmap_headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[col_letter].width = width

    # Validations
    dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked,Deferred"', allow_blank=True)
    dv_area = DataValidation(type="list", formula1='"A.8.5 Authentication,A.8.5 MFA,A.8.2 Privileged Access,A.8.2 Monitoring,A.8.3 Access Restriction"', allow_blank=True)

    ws.add_data_validation(dv_priority)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_area)

    # Pre-populate 200 data rows
    for i in range(200):
        data_row = row + 1 + i
        ws[f"A{data_row}"] = f"REM-{i+1:03d}" if i < 5 else ""
        ws[f"A{data_row}"].font = Font(color="808080") if i >= 5 else Font()

        # Apply input formatting
        for col in ["B", "C", "F", "G", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
            ws[f"{col}{data_row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{data_row}"].border = styles["border"]

        # Apply validations
        dv_priority.add(ws[f"D{data_row}"])
        dv_area.add(ws[f"E{data_row}"])
        dv_status.add(ws[f"H{data_row}"])

        ws[f"D{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{data_row}"].border = styles["border"]
        ws[f"E{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{data_row}"].border = styles["border"]
        ws[f"H{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{data_row}"].border = styles["border"]

    ws.freeze_panes = "A37"


# ============================================================================
# SECTION 6: KPIs & METRICS
# ============================================================================

def create_kpis_metrics(ws, styles):
    """Create detailed KPIs & Metrics tracking sheet."""

    # Header
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        "KEY PERFORMANCE INDICATORS & METRICS - AUTHENTICATION & ACCESS\n"
        "Detailed tracking of security metrics and compliance targets"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # KPI Categories
    kpi_categories = [
        {
            "name": "AUTHENTICATION METRICS (A.8.5)",
            "kpis": [
                ("MFA Enrollment - All Users", "%", "", "≥85%", ""),
                ("MFA Enrollment - Privileged Users", "%", "", "100%", ""),
                ("MFA Enrollment - Remote Access", "%", "", "100%", ""),
                ("MFA Enrollment - Contractors", "%", "", "100%", ""),
                ("SSO Adoption Rate", "%", "", "≥80%", ""),
                ("Password Policy Compliance", "%", "", "100%", ""),
                ("Failed Login Attempts (Monthly Avg)", "Count", "", "<100", ""),
                ("Account Lockouts (Monthly)", "Count", "", "<50", ""),
            ]
        },
        {
            "name": "PRIVILEGED ACCESS METRICS (A.8.2)",
            "kpis": [
                ("Shared Accounts in PAM Vault", "%", "", "100%", ""),
                ("Service Accounts in PAM Vault", "%", "", "≥95%", ""),
                ("Session Recording Coverage - Tier 0", "%", "", "100%", ""),
                ("Session Recording Coverage - Tier 1", "%", "", "≥95%", ""),
                ("Tier Isolation Violations", "Count", "", "0", ""),
                ("Privileged Access Reviews - On Time", "%", "", "100%", ""),
                ("Emergency Access Requests (Monthly)", "Count", "", "<10", ""),
                ("Orphaned Privileged Accounts", "Count", "", "0", ""),
            ]
        },
        {
            "name": "ACCESS RESTRICTION METRICS (A.8.3)",
            "kpis": [
                ("Default Deny Policy Compliance", "%", "", "≥95%", ""),
                ("Excessive Permission Violations", "Count", "", "<20", ""),
                ("Access Reviews Completed On Time", "%", "", "100%", ""),
                ("Segregation of Duties Violations", "Count", "", "0", ""),
                ("Data Classification Compliance", "%", "", "100%", ""),
            ]
        },
    ]

    row = 3
    for category in kpi_categories:
        # Category header
        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = category["name"]
        ws[f"A{row}"].font = styles["subheader"]["font"]
        ws[f"A{row}"].fill = styles["subheader"]["fill"]
        ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

        row += 1
        # Column headers
        kpi_headers = ["KPI Name", "Unit", "Current Value", "Target", "Status", "Q1", "Q2", "Q3", "Q4", "YTD Trend", "Owner", "Notes"]
        for col_idx, header in enumerate(kpi_headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = styles["column_header"]["font"]
            cell.fill = styles["column_header"]["fill"]
            cell.alignment = styles["column_header"]["alignment"]
            cell.border = styles["column_header"]["border"]

        row += 1
        # KPI data rows
        for kpi_name, unit, current, target, status in category["kpis"]:
            ws.cell(row=row, column=1, value=kpi_name).border = styles["border"]
            ws.cell(row=row, column=2, value=unit).border = styles["border"]
            ws.cell(row=row, column=3, value=current)
            ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
            ws.cell(row=row, column=3).border = styles["border"]
            ws.cell(row=row, column=4, value=target).border = styles["border"]
            ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
            ws.cell(row=row, column=5).border = styles["border"]

            # Quarterly columns
            for col in range(6, 10):
                ws.cell(row=row, column=col).fill = styles["input_cell"]["fill"]
                ws.cell(row=row, column=col).border = styles["border"]

            ws.cell(row=row, column=10).fill = styles["input_cell"]["fill"]
            ws.cell(row=row, column=10).border = styles["border"]
            ws.cell(row=row, column=11).fill = styles["input_cell"]["fill"]
            ws.cell(row=row, column=11).border = styles["border"]
            ws.cell(row=row, column=12).fill = styles["input_cell"]["fill"]
            ws.cell(row=row, column=12).border = styles["border"]

            row += 1

        row += 1  # Space between categories

    # Set column widths
    widths = [35, 8, 14, 12, 12, 12, 12, 12, 12, 12, 18, 25]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"On Target,At Risk,Off Target,N/A"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"E4:E{row}")

    # Trend dropdown
    dv_trend = DataValidation(type="list", formula1='"↑ Improving,→ Stable,↓ Declining"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add(f"J4:J{row}")

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create evidence register with 100 data rows."""

    # Header
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        "EVIDENCE REGISTER - AUTHENTICATION & ACCESS CONTROLS\n"
        "Audit evidence index linking to source documents and artifacts"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Summary
    ws["A3"] = "EVIDENCE SUMMARY"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    ws["A4"] = "Total Evidence Items:"
    ws["B4"] = "=COUNTA(A7:A106)-COUNTBLANK(A7:A106)"
    ws["B4"].font = Font(bold=True, color="0000FF")
    ws["A5"] = "Evidence by Status:"
    ws["B5"] = "Current: " + "=COUNTIF(K7:K106,\"Current\")"

    # Column Headers
    row = 6
    evidence_headers = [
        ("A", "Evidence ID", 14),
        ("B", "Control Reference", 18),
        ("C", "Evidence Type", 18),
        ("D", "Description", 40),
        ("E", "Source System", 20),
        ("F", "File Location/Path", 35),
        ("G", "Collected By", 18),
        ("H", "Collection Date", 14),
        ("I", "Review Date", 14),
        ("J", "Reviewer", 18),
        ("K", "Status", 14),
        ("L", "Notes", 30),
    ]

    for col_letter, header, width in evidence_headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[col_letter].width = width

    # Validations
    dv_type = DataValidation(type="list", formula1='"Screenshot,Configuration Export,Log Extract,Report,Policy Document,Procedure,Test Result,Audit Finding,Risk Assessment,Other"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Current,Requires Update,Archived,Missing"', allow_blank=True)
    dv_control = DataValidation(type="list", formula1='"A.8.2,A.8.3,A.8.5,Multiple"', allow_blank=True)

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_control)

    # Pre-populate 100 data rows
    for i in range(100):
        data_row = row + 1 + i
        ws[f"A{data_row}"] = f"EV-{i+1:03d}" if i < 5 else ""
        ws[f"A{data_row}"].font = Font(color="808080") if i >= 5 else Font()

        for col in ["B", "D", "E", "F", "G", "H", "I", "J", "L"]:
            ws[f"{col}{data_row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{data_row}"].border = styles["border"]

        dv_control.add(ws[f"B{data_row}"])
        dv_type.add(ws[f"C{data_row}"])
        dv_status.add(ws[f"K{data_row}"])

        ws[f"C{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{data_row}"].border = styles["border"]
        ws[f"K{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"K{data_row}"].border = styles["border"]

    ws.freeze_panes = "A7"


# ============================================================================
# SECTION 8: ACTION ITEMS & FOLLOW-UP
# ============================================================================

def create_action_items(ws, styles):
    """Create action items tracking sheet with 100 data rows."""

    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = (
        "ACTION ITEMS & FOLLOW-UP TRACKER\n"
        "Open actions requiring completion for compliance"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Summary
    ws["A3"] = "ACTION SUMMARY"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    ws["A4"] = "Total Actions:"
    ws["B4"] = "=COUNTA(A8:A107)-COUNTBLANK(A8:A107)"
    ws["B4"].font = Font(bold=True, color="0000FF")
    ws["A5"] = "Open:"
    ws["B5"] = "=COUNTIF(I8:I107,\"Open\")+COUNTIF(I8:I107,\"In Progress\")"
    ws["B5"].font = Font(color="0000FF")
    ws["A6"] = "Overdue:"
    ws["B6"] = "=COUNTIFS(I8:I107,\"<>Completed\",G8:G107,\"<\"&TODAY())"
    ws["B6"].font = Font(color="C00000")

    # Column Headers
    row = 7
    action_headers = [
        ("A", "Action ID", 12),
        ("B", "Source", 16),
        ("C", "Description", 45),
        ("D", "Owner", 18),
        ("E", "Priority", 12),
        ("F", "Created Date", 14),
        ("G", "Due Date", 14),
        ("H", "Completed Date", 14),
        ("I", "Status", 14),
        ("J", "Comments", 35),
        ("K", "Evidence Ref", 14),
    ]

    for col_letter, header, width in action_headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[col_letter].width = width

    # Validations
    dv_source = DataValidation(type="list", formula1='"Gap Analysis,Risk Register,Audit Finding,Management Review,Incident,Other"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Open,In Progress,Completed,Cancelled,Deferred"', allow_blank=True)

    ws.add_data_validation(dv_source)
    ws.add_data_validation(dv_priority)
    ws.add_data_validation(dv_status)

    # Pre-populate 100 data rows
    for i in range(100):
        data_row = row + 1 + i
        ws[f"A{data_row}"] = f"ACT-{i+1:03d}" if i < 5 else ""
        ws[f"A{data_row}"].font = Font(color="808080") if i >= 5 else Font()

        for col in ["C", "D", "F", "G", "H", "J", "K"]:
            ws[f"{col}{data_row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{data_row}"].border = styles["border"]

        dv_source.add(ws[f"B{data_row}"])
        dv_priority.add(ws[f"E{data_row}"])
        dv_status.add(ws[f"I{data_row}"])

        ws[f"B{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{data_row}"].border = styles["border"]
        ws[f"E{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{data_row}"].border = styles["border"]
        ws[f"I{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"I{data_row}"].border = styles["border"]

    ws.freeze_panes = "A8"


# ============================================================================
# SECTION 9: AUDIT & COMPLIANCE LOG
# ============================================================================

def create_audit_log(ws, styles):
    """Create audit and compliance log sheet."""

    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = (
        "AUDIT & COMPLIANCE LOG\n"
        "Record of audits, assessments, and compliance activities"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Column Headers
    row = 3
    audit_headers = [
        ("A", "Entry ID", 12),
        ("B", "Date", 14),
        ("C", "Activity Type", 18),
        ("D", "Description", 45),
        ("E", "Conducted By", 20),
        ("F", "Findings Summary", 35),
        ("G", "Actions Required", 30),
        ("H", "Status", 14),
        ("I", "Follow-up Date", 14),
        ("J", "Notes", 30),
    ]

    for col_letter, header, width in audit_headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[col_letter].width = width

    # Validations
    dv_type = DataValidation(type="list", formula1='"Internal Audit,External Audit,Self-Assessment,Management Review,Surveillance Audit,Certification Audit,Gap Assessment,Other"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Scheduled,In Progress,Completed,Findings Open,Findings Closed"', allow_blank=True)

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_status)

    # Pre-populate 50 data rows
    for i in range(50):
        data_row = row + 1 + i
        ws[f"A{data_row}"] = f"AUD-{i+1:03d}" if i < 3 else ""
        ws[f"A{data_row}"].font = Font(color="808080") if i >= 3 else Font()

        for col in ["B", "D", "E", "F", "G", "I", "J"]:
            ws[f"{col}{data_row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{data_row}"].border = styles["border"]

        dv_type.add(ws[f"C{data_row}"])
        dv_status.add(ws[f"H{data_row}"])

        ws[f"C{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{data_row}"].border = styles["border"]
        ws[f"H{data_row}"].fill = styles["input_cell"]["fill"]
        ws[f"H{data_row}"].border = styles["border"]

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create approval and sign-off workflow sheet."""

    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = (
        "DASHBOARD APPROVAL & SIGN-OFF\n"
        "Executive review and approval workflow"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Document Summary
    ws["A3"] = "DOCUMENT SUMMARY"
    ws["A3"].font = Font(bold=True, size=11)

    row = 4
    summary_items = [
        ("Dashboard Version:", "1.0"),
        ("Assessment Period:", ""),
        ("Overall Compliance Score:", "='Executive Dashboard'!B17"),
        ("Total Critical Gaps:", "='Executive Dashboard'!B18"),
        ("Total High-Risk Items:", "='Executive Dashboard'!B19"),
        ("Evidence Items Collected:", "='Evidence Register'!B4"),
        ("Open Action Items:", "='Action Items & Follow-up'!B5"),
    ]

    for label, value in summary_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        elif "=" in str(value):
            ws[f"B{row}"].font = Font(color="0000FF")
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approval Workflow
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "APPROVAL WORKFLOW"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 1
    approval_headers = ["Level", "Role", "Name", "Signature", "Date", "Comments"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    approval_roles = [
        ("1", "Prepared By", "Security Analyst / Assessor"),
        ("2", "Reviewed By", "IT Security Manager"),
        ("3", "Approved By", "CISO / Head of IT Security"),
        ("4", "Noted By", "CIO / IT Director"),
    ]

    row += 1
    for level, role_type, role_name in approval_roles:
        ws.cell(row=row, column=1, value=level).border = styles["border"]
        ws.cell(row=row, column=2, value=role_type).border = styles["border"]
        ws.cell(row=row, column=3, value=role_name)
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=3).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).border = styles["border"]
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6).border = styles["border"]
        row += 1

    # Next Review
    row += 2
    ws[f"A{row}"] = "NEXT REVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    ws[f"A{row}"] = "Next Review Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]

    row += 1
    ws[f"A{row}"] = "Review Frequency:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "Quarterly"
    ws[f"B{row}"].border = styles["border"]

    # Column widths
    widths = [15, 20, 25, 20, 15, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Generate the compliance dashboard workbook."""
    logger.info("\n" + "=" * 80)
    logger.info("ISMS-IMP-A.8.2-3-5.6 - Authentication & Privileged Access Compliance Dashboard")
    logger.info("ISO/IEC 27001:2022 - Controls A.8.2, A.8.3, A.8.5")
    logger.info("=" * 80 + "\n")

    wb = create_workbook()
    styles = setup_styles()

    logger.info("Creating dashboard sheets...")

    logger.info("  [1/9] Creating Executive Dashboard...")
    create_executive_dashboard(wb["Executive Dashboard"], styles)

    logger.info("  [2/9] Creating Gap Analysis (200 rows)...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("  [3/9] Creating Risk Register (100 rows)...")
    create_risk_register(wb["Risk Register"], styles)

    logger.info("  [4/9] Creating Remediation Roadmap (200 rows)...")
    create_remediation_roadmap(wb["Remediation Roadmap"], styles)

    logger.info("  [5/9] Creating KPIs & Metrics...")
    create_kpis_metrics(wb["KPIs & Metrics"], styles)

    logger.info("  [6/9] Creating Evidence Register (100 rows)...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("  [7/9] Creating Action Items & Follow-up (100 rows)...")
    create_action_items(wb["Action Items & Follow-up"], styles)

    logger.info("  [8/9] Creating Audit & Compliance Log...")
    create_audit_log(wb["Audit & Compliance Log"], styles)

    logger.info("  [9/9] Creating Approval Sign-Off...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.8.2-3-5.6_Compliance_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info("\n" + "=" * 80)
    logger.info(f"{CHECK} SUCCESS: {filename}")
    logger.info("=" * 80)
    logger.info("\nGenerated Dashboard Structure:")
    logger.info("  1. Executive Dashboard - Overall compliance status, KPIs, critical gaps")
    logger.info("  2. Gap Analysis - 200 rows for consolidated gap tracking")
    logger.info("  3. Risk Register - 100 rows for risk management")
    logger.info("  4. Remediation Roadmap - 200 rows with timeline/budget tracking")
    logger.info("  5. KPIs & Metrics - Detailed KPI tracking by control area")
    logger.info("  6. Evidence Register - 100 rows for audit evidence")
    logger.info("  7. Action Items & Follow-up - 100 rows for action tracking")
    logger.info("  8. Audit & Compliance Log - Audit trail")
    logger.info("  9. Approval Sign-Off - 4-level approval workflow")
    logger.info("\nNext Steps:")
    logger.info("  1. Place dashboard in same folder as normalized assessment workbooks")
    logger.info("  2. Open dashboard in Excel")
    logger.info("  3. Click 'Update Links' when prompted to connect source workbooks")
    logger.info("  4. Review Executive Dashboard for overall compliance status")
    logger.info("  5. Complete Gap Analysis with findings from assessments")
    logger.info("  6. Update Risk Register with identified risks")
    logger.info("  7. Create Remediation Roadmap items")
    logger.info("  8. Obtain executive approvals")
    logger.info("=" * 80 + "\n")

    return filename


if __name__ == "__main__":
    try:
        main()
        sys.exit(0)
    except Exception as e:
        logger.error(f"\n{XMARK} ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
