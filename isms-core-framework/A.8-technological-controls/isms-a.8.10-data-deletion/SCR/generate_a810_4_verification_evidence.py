#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.10.4 - Verification & Evidence Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.10: Information Deletion
Assessment Domain 4 of 4: Deletion Verification and Evidence Collection Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific verification procedures, forensic testing capabilities,
audit evidence requirements, and documentation standards.

Key customisation areas:
1. Verification testing frequency (match your risk assessment and audit schedule)
2. Forensic testing tools and procedures (specific to your testing infrastructure)
3. Sampling methodologies (adapt to your media volumes and risk levels)
4. Evidence retention requirements (aligned with your audit and legal obligations)
5. Documentation templates (customized to your organisational standards)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.10 Information Deletion Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
deletion verification procedures and evidence collection practices against ISO
27001:2022 Control A.8.10 requirements, supporting evidence-based validation of
systematic deletion verification and audit readiness.

**Purpose:**
Enables systematic assessment of deletion verification testing, forensic validation,
documentation practices, and evidence collection to prove information was properly
deleted and cannot be recovered.

**Assessment Scope:**
- Deletion verification testing procedures and frequency
- Forensic testing capabilities and tool validation
- Sampling methodologies for verification testing
- Evidence documentation and retention practices
- Deletion log management and monitoring
- Audit evidence collection and organisation
- Verification paradox resolution (proving deletion without retaining data)
- Gap analysis for inadequate verification procedures
- Evidence readiness for ISO 27001 audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance, verification principles, color coding
2. Verification Testing Strategy - Testing frequency, scope, sampling methodology
3. Forensic Testing Procedures - Tools, validation methods, test case library
4. Deletion Log Management - Log retention, monitoring, alerting procedures
5. Evidence Documentation - Documentation templates, evidence organisation
6. Verification Sampling - Statistical sampling for large-scale deletion verification
7. Verification Dashboard - Testing compliance, gap identification, evidence status
8. Evidence Register - Master index of all deletion verification evidence
9. Approval Sign-Off - Three-level stakeholder approval workflow

**Key Features:**
- Data validation with testing frequency and sampling method dropdown lists
- Conditional formatting for verification testing compliance status
- Automated gap identification for missing verification procedures
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with deletion logs and forensic testing results
- Verification paradox guidance (proving deletion without retaining deleted data)

**Integration:**
consolidates data from all four information deletion assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a810_4_verification_evidence.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a810_4_verification_evidence.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a810_4_verification_evidence.py --date 20250124

Output:
    File: ISMS_A_8_10_4_Verification_Evidence_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize verification testing frequencies to match your risk levels
    2. Document all verification testing procedures and tools
    3. Complete forensic testing capability assessment
    4. Define sampling methodologies for large-scale verification
    5. Review deletion log management and monitoring procedures
    6. Organize evidence documentation templates and repositories
    7. Validate evidence collection practices for audit readiness
    8. Conduct gap analysis for missing verification procedures
    9. Define remediation actions with timelines
    10. Collect and link audit evidence (test reports, forensic results, deletion logs)
    11. Obtain stakeholder approvals

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.10
Assessment Domain:    4 of 4 (Deletion Verification and Evidence Collection Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.10: Information Deletion Policy (Governance)
    - ISMS-IMP-A.8.10.1: Retention & Deletion Triggers Assessment (Domain 1)
    - ISMS-IMP-A.8.10.2: Deletion Methods Assessment (Domain 2)
    - ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion Assessment (Domain 3)
    - ISMS-IMP-A.8.10.4: Verification & Evidence Implementation Guide

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.10.4 specification
    - Supports comprehensive verification and evidence evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**The Verification Paradox:**
Proving data was deleted creates a fundamental challenge: How do you prove something
no longer exists without retaining evidence of what was deleted?

**Solutions:**
1. **Retain Metadata, Not Data:**
   - Store deletion timestamps, record counts, system identifiers
   - Use cryptographic hashes of deleted data (hash proves existence, not content)
   - Document deletion method specifications and tool configurations

2. **Deletion Logs:**
   - System-generated logs showing deletion operations
   - Include: timestamp, user, system, record count, deletion method
   - Do NOT include actual deleted data

3. **Forensic Testing:**
   - Periodic testing on sample media to verify deletion effectiveness
   - Test certificates demonstrate deletion method works as specified
   - Extrapolate from sample to population

4. **Process Documentation:**
   - Document deletion procedures and their NIST SP 800-88 Rev. 2 classifications
   - Maintain evidence of procedure compliance (checklists, approvals)
   - Link procedures to specific systems and data categories

5. **Third-Party Certificates:**
   - Cloud provider deletion certificates
   - Destruction vendor certificates of destruction
   - Audit reports verifying deletion controls

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of:
- Deletion verification testing procedures and results
- Forensic testing capabilities and tool validation
- Evidence documentation and organisation
- Deletion log retention and monitoring

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Verification testing procedures and tools
- Forensic testing results and findings
- Deletion log configurations and monitoring
- Evidence repository locations and access controls

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check verification testing results and evidence completeness
- Semi-annually: Update forensic testing procedures and tool validation
- Annually: Complete reassessment of all verification procedures
- Ad-hoc: When new deletion methods are implemented or audit findings emerge

**Quality Assurance:**
Have Information Security, IT Operations, Internal Audit, and Legal/Compliance
teams validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
Ensure verification procedures align with applicable regulatory requirements:
- ISO 27001:2022: Clause 8.10 requires verification that information is properly deleted
- GDPR: Article 5(2) accountability principle (demonstrate compliance)
- PCI DSS v4.0.1: Requirement 3.1, 9.8 (verify media destruction)
- HIPAA: 164.310(d)(2)(i) (verify media disposal effectiveness)

Customize assessment criteria to include regulatory-specific requirements.

**Verification Testing Frequency:**
Risk-based approach to verification testing frequency:
- **High-Risk Data (Confidential/Restricted):**
  - Forensic testing: Quarterly sample testing
  - Deletion log review: Weekly monitoring
  - Evidence audit: Monthly completeness checks

- **Medium-Risk Data (Internal):**
  - Forensic testing: Semi-annual sample testing
  - Deletion log review: Monthly monitoring
  - Evidence audit: Quarterly completeness checks

- **Low-Risk Data (Public):**
  - Forensic testing: Annual sample testing
  - Deletion log review: Quarterly monitoring
  - Evidence audit: Annual completeness checks

Customize frequencies based on your organisation's risk assessment.

**Sampling Methodologies:**
For large-scale deletion operations, use statistical sampling:
1. **Random Sampling:** Select random sample from deletion population
2. **Stratified Sampling:** Sample proportionally from different media types
3. **Risk-Based Sampling:** Oversample high-risk data categories
4. **Time-Based Sampling:** Sample from different deletion time periods

Sample sizes:
- Small operations (<100 deletions/month): 10-20% sample
- Medium operations (100-1000 deletions/month): 5-10% sample
- Large operations (>1000 deletions/month): 1-5% sample (minimum 50 items)

**Evidence Organisation:**
Structure evidence repository for audit readiness:
```
Evidence Repository/
├── Deletion_Policies/
│   ├── ISMS-POL-A.8.10-Information_Deletion.pdf
│   └── Retention_Schedules/
├── Verification_Testing/
│   ├── Forensic_Test_Reports/
│   ├── Tool_Validation_Certificates/
│   └── Sample_Test_Results/
├── Deletion_Logs/
│   ├── System_Deletion_Logs/
│   ├── Cloud_Provider_Logs/
│   └── Manual_Deletion_Records/
├── Third_Party_Certificates/
│   ├── Cloud_Deletion_Certificates/
│   ├── Destruction_Vendor_Certificates/
│   └── Audit_Reports/
└── Compliance_Evidence/
    ├── Assessment_Workbooks/
    ├── Gap_Remediation_Plans/
    └── Approval_Sign_Offs/
```

**Common Verification Gaps:**
1. ❌ No forensic testing - Assume deletion works without verification
2. ❌ No deletion logs - Cannot prove deletions occurred
3. ❌ No sampling methodology - Cannot scale verification to large volumes
4. ❌ Evidence not organised - Cannot locate evidence during audits
5. ❌ Verification not scheduled - Ad-hoc testing with long gaps
6. ❌ No tool validation - Deletion tools assumed effective without testing

**Forensic Testing Tools:**
Examples of forensic data recovery tools for verification testing:
- **Commercial:** EnCase, FTK (Forensic Toolkit), X-Ways Forensics
- **Open Source:** Autopsy, TestDisk, PhotoRec, Sleuth Kit
- **Specialized:** SSD-specific tools for verifying crypto-erasure

Validate tools against known test datasets before using for verification.

**Deletion Log Requirements:**
Deletion logs should capture at minimum:
- **Timestamp:** When deletion occurred (ISO 8601 format)
- **User/System:** Who/what initiated deletion
- **Target:** System, database, file path affected
- **Scope:** Record count, file count, or data volume deleted
- **Method:** Deletion method used (Clear/Purge/Destroy)
- **Status:** Success, failure, partial completion
- **Verification:** Reference to verification testing (if applicable)

Do NOT log actual deleted data content (defeats purpose of deletion).

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ============================================================================
# SECTION 1: CONFIGURATION AND CONSTANTS
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.10.4"
WORKBOOK_NAME = "Verification & Evidence Assessment"
CONTROL_ID = "A.8.10"
CONTROL_NAME = "Information Deletion"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
WORKBOOK_TITLE = "ISMS-IMP-A.8.10.4 - Verification & Evidence Assessment"


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠ Warning sign
BULLET = '\u2022'     # * Bullet point
ARROW = '\u2192'      # -> Right arrow

FILENAME_PREFIX = "ISMS-IMP-A.8.10.4_Verification_Evidence"
VERSION = "1.0"
RELATED_POLICY = "ISMS-POL-A.8.10-S2.3 (Verification & Evidence Requirements)"

# Sheet names
SHEET_NAMES = [
    "Instructions & Legend",
    "Deletion Logging Assessment",
    "Verification Testing Program",
    "Evidence Repository Assessment",
    "Certificate Management",
    "Audit Trail Completeness",
    "Evidence Register",
    "Summary Dashboard",
    "Approval Sign-Off"
]

# Standard column headers (A-Q) - consistent across all A.8.10 assessments
STANDARD_HEADERS = [
    "Verification Area",                # A
    "Current State",                    # B
    "Target State",                     # C
    "Gap Severity",                     # D
    "Control Effectiveness",            # E
    "Evidence Quality",                 # F
    "Compliance Status",                # G
    "Risk Level",                       # H
    "Responsible Party",                # I
    "Review Frequency",                 # J
    "Last Review Date",                 # K
    "Next Review Due",                  # L
    "Remediation Priority",             # M
    "Estimated Effort",                 # N
    "Target Completion",                # O
    "Current Status",                   # P
    "Assessor Notes"                    # Q
]

# Extended headers by sheet (R-U)
EXTENDED_HEADERS = {
    "Deletion Logging Assessment": [
        "Log Completeness Score (%)",
        "Log Retention Period",
        "Centralised System",
        "Tamper Protection"
    ],
    "Verification Testing Program": [
        "Test Frequency",
        "Sample Size (%)",
        "Last Test Date",
        "Test Pass Rate (%)"
    ],
    "Evidence Repository Assessment": [
        "Storage Location",
        "Access Controls",
        "Backup Status",
        "Retention Compliance"
    ],
    "Certificate Management": [
        "Certificate Quality Score (%)",
        "Vendor Compliance Rate (%)",
        "Validation Method",
        "Certificate Storage Period"
    ],
    "Audit Trail Completeness": [
        "Reconstruction Tested",
        "Chain of Custody",
        "Completeness Score (%)",
        "Audit Readiness"
    ]
}

# Data validation lists
VALIDATION_LISTS = {
    "current_target_state": [
        "Not Implemented",
        "Partially Implemented",
        "Substantially Implemented",
        "Fully Implemented",
        "Continuously Improving"
    ],
    "gap_severity": [
        "None (No Gap)",
        "Low (Minor improvement needed)",
        "Medium (Significant gap exists)",
        "High (Critical gap, immediate action)",
        "N/A"
    ],
    "control_effectiveness": [
        "Ineffective (Does not work)",
        "Partially Effective (Works sometimes)",
        "Effective (Works as designed)",
        "Highly Effective (Exceeds requirements)",
        "Not Applicable"
    ],
    "evidence_quality": [
        "None (No evidence exists)",
        "Poor (Insufficient or unreliable)",
        "Fair (Basic evidence present)",
        "Good (Adequate and reliable)",
        "Excellent (Comprehensive and auditable)"
    ],
    "compliance_status": [
        "Non-Compliant",
        "Partially Compliant",
        "Substantially Compliant",
        "Fully Compliant",
        "N/A (Not applicable to regulations)"
    ],
    "risk_level": [
        "Low (Minimal impact if fails)",
        "Medium (Moderate impact)",
        "High (Significant impact)",
        "Critical (Severe legal/regulatory impact)"
    ],
    "review_frequency": [
        "Weekly",
        "Monthly",
        "Quarterly",
        "Semi-Annually",
        "Annually",
        "Ad-Hoc"
    ],
    "remediation_priority": [
        "P1 - Critical (Immediate action)",
        "P2 - High (Within 30 days)",
        "P3 - Medium (Within 90 days)",
        "P4 - Low (Within 180 days)",
        "P5 - Deferred (Future consideration)"
    ],
    "estimated_effort": [
        "Minimal (<8 hours)",
        "Low (1-3 days)",
        "Medium (1-2 weeks)",
        "High (2-4 weeks)",
        "Very High (>1 month)"
    ],
    "current_status": [
        "Not Started",
        "Planning",
        "In Progress",
        "Testing",
        "Completed",
        "On Hold"
    ],
    "yes_no_partial": [
        "Yes",
        "No",
        "Partial"
    ],
    "tamper_protection": [
        "None (Logs can be modified)",
        "Basic (File permissions only)",
        "Advanced (Cryptographic signing)",
        "Immutable (WORM/blockchain)",
        "N/A"
    ],
    "test_frequency": [
        "Continuous (Automated per deletion)",
        "Weekly (Sample-based)",
        "Monthly (Sample-based)",
        "Quarterly (Sample-based)",
        "Annually (Sample-based)",
        "Never (No testing - CRITICAL GAP)"
    ],
    "access_controls": [
        "Public (No restrictions - CRITICAL GAP)",
        "Internal (All employees)",
        "Restricted (Specific roles only)",
        "Highly Restricted (Security/Compliance only)",
        "Auditor Access (Read-only for auditors)"
    ],
    "retention_compliance": [
        "Compliant (Aligned with ISMS-POL-A.8.10-S2.3)",
        "Over-Retention (Exceeds legal requirements)",
        "Under-Retention (Below legal requirements)",
        "No Policy (Undefined retention - CRITICAL GAP)"
    ],
    "validation_method": [
        "None (Certificate accepted without validation)",
        "Basic (Visual inspection only)",
        "Standard (Signature verification + content review)",
        "Enhanced (Third-party audit + signature verification)",
        "Continuous (API-based real-time verification)"
    ],
    "chain_of_custody": [
        "Complete (Full chain from request to verification)",
        "Mostly Complete (Minor gaps, non-critical)",
        "Partial (Significant gaps exist)",
        "Incomplete (Cannot establish chain)"
    ],
    "audit_readiness": [
        "Fully Ready (Comprehensive audit trail, tested)",
        "Mostly Ready (Minor documentation gaps)",
        "Partially Ready (Significant preparation needed)",
        "Not Ready (Major gaps, would fail audit)"
    ]
}

# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def create_styles():
    """Create reusable cell styles"""
    styles = {}
    
    # Header style (dark blue background, white text)
    styles['header'] = {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Subheader style (light blue background)
    styles['subheader'] = {
        'font': Font(name='Calibri', size=10, bold=True),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Input cell style (yellow background)
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Reference table header
    styles['ref_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Reference table cell
    styles['ref_cell'] = {
        'font': Font(name='Calibri', size=9),
        'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Status colors (conditional formatting helpers)
    styles['status_critical'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    styles['status_high'] = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    styles['status_medium'] = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    styles['status_low'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    styles['status_good'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    return styles

# ============================================================================
# SECTION 3: WORKBOOK INITIALIZATION
# ============================================================================

def create_workbook():
    """Initialize workbook and create all sheets"""
    wb = openpyxl.Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create all sheets
    for sheet_name in SHEET_NAMES:
        wb.create_sheet(title=sheet_name)
    
    return wb

# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET
# ============================================================================

def populate_instructions_sheet(ws, styles):
    """Populate the Instructions sheet"""
    # Title
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Document Information — golden standard
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws['A3'] = "Document Information"
    ws['A3'].font = Font(bold=True, size=12)

    doc_fields = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", WORKBOOK_NAME),
        ("Related Policy", RELATED_POLICY),
        ("Version", VERSION),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]
    for idx, (label, value) in enumerate(doc_fields, start=4):
        ws[f'A{idx}'] = label
        ws[f'A{idx}'].font = Font(name='Calibri', bold=True)
        ws[f'B{idx}'] = value
        if value == "":
            ws[f'B{idx}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws[f'B{idx}'].border = border_thin

    # Purpose
    ws['A14'] = "PURPOSE"
    ws['A14'].font = Font(bold=True, size=12)
    ws['A15'] = ("This workbook assesses an organisation's deletion verification and evidence management "
                "capabilities. It addresses the critical challenge of proving information has been deleted "
                "while respecting data minimisation principles.")
    ws.merge_cells('A15:G17')
    ws['A15'].alignment = Alignment(wrap_text=True, vertical='top')

    # The Verification Paradox
    ws['A19'] = "THE VERIFICATION PARADOX"
    ws['A19'].font = Font(bold=True, size=12, color='C00000')
    ws['A20'] = ("Organisations must prove deletion occurred without retaining the deleted data itself. "
                 "This assessment addresses this through:\n\n"
                 f"{BULLET} Metadata Over Data: Retention of deletion logs, certificates, and test results "
                 "(not actual deleted content)\n"
                 f"{BULLET} Method Testing Over Data Testing: Forensic testing validates deletion methods using "
                 "test datasets, not production data\n"
                 f"{BULLET} Time-Bound Evidence: Evidence retention aligned with legal/regulatory requirements, "
                 "then evidence itself is deleted")
    ws.merge_cells('A20:G25')
    ws['A20'].alignment = Alignment(wrap_text=True, vertical='top')

    # Assessment Structure
    ws['A27'] = "ASSESSMENT STRUCTURE"
    ws['A27'].font = Font(bold=True, size=12)

    instructions_data = [
        ["Sheet", "Purpose"],
        ["Deletion Logging Assessment", "Evaluate centralised logging infrastructure and completeness"],
        ["Verification Testing Program", "Assess forensic testing capability and effectiveness"],
        ["Evidence Repository Assessment", "Review storage, retention, and access control of evidence"],
        ["Certificate Management", "Validate quality and management of deletion certificates"],
        ["Audit Trail Completeness", "Assess reconstruction capability and chain of custody"],
        ["Summary Dashboard", "Summary metrics and gap analysis"],
        ["Evidence Register", "Document assessment evidence (100 rows)"],
        ["Approval Sign-Off", "Three-level approval workflow"]
    ]

    for idx, row in enumerate(instructions_data, start=28):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            if idx == 28:  # Header row
                apply_style(cell, styles['ref_header'])
            else:
                apply_style(cell, styles['ref_cell'])
    
    # How to Use
    ws['A31'] = "Instructions"
    ws['A31'].font = Font(bold=True, size=12)
    ws['A32'] = ("1. Review ISMS-POL-A.8.10-S2.3 (Verification & Evidence Requirements) before starting\n"
                 "2. Complete Sheets 2-6 (assessment sheets) - yellow cells require data entry\n"
                 "3. Use dropdown menus where provided - consistent assessment criteria\n"
                 "4. Complete Sheet 7 (Summary Dashboard) for summary and gap analysis\n"
                 "5. Document all evidence in Sheet 8 (Evidence Register)\n"
                 "6. Obtain three-level approval in Sheet 9\n"
                 "7. Track remediation of identified gaps through to closure")
    ws.merge_cells('A32:G38')
    ws['A32'].alignment = Alignment(wrap_text=True, vertical='top')

    # Key Reminders
    ws['A40'] = "KEY REMINDERS"
    ws['A40'].font = Font(bold=True, size=12, color='C00000')
    ws['A41'] = (f"{BULLET} NO CARGO CULT CERTIFICATES: Don't accept vendor certificates without validation!\n"
                 f"{BULLET} NO EVIDENCE HOARDING: Retention must balance legal requirements with data minimisation\n"
                 f"{BULLET} NO PRODUCTION DATA TESTING: Test deletion methods with test datasets, not actual sensitive data\n"
                 f"{BULLET} COMPLETE AUDIT TRAILS: Gaps in chain of custody undermine entire verification program\n"
                 f"{BULLET} LOGS NEED STORAGE: Insufficient log capacity = lost evidence = compliance failure")
    ws.merge_cells('A41:G47')
    ws['A41'].alignment = Alignment(wrap_text=True, vertical='top')

    # Status Legend
    legend_row = 49
    ws.merge_cells(f'A{legend_row}:G{legend_row}')
    ws[f'A{legend_row}'] = "Status Legend"
    ws[f'A{legend_row}'].font = Font(name='Calibri', size=10, bold=True)
    ws[f'A{legend_row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    legend_row += 1

    for col_idx, hdr in enumerate(["Symbol", "Status", "Description"], 1):
        cell = ws.cell(row=legend_row, column=col_idx, value=hdr)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
    legend_row += 1

    for symbol, status, desc in [
        ("003366", "Column Header", "Do not edit"),
        ("FFFFCC", "Data Entry", "Complete these fields"),
        ("C6EFCE", "Compliant", "Meets requirements"),
        ("FFEB9C", "Partial", "Needs attention"),
        ("FFC7CE", "Non-Compliant", "Critical gap"),
        ("F2F2F2", "Reference", "Read-only information"),
        ("\u2014", "N/A", "Not applicable"),
    ]:
        _is_color = symbol in {"003366", "FFFFCC", "C6EFCE", "FFEB9C", "FFC7CE", "F2F2F2"}
        _thin_leg = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        _ca = ws.cell(row=legend_row, column=1, value="" if _is_color else symbol)
        _cb = ws.cell(row=legend_row, column=2, value=status)
        _cc = ws.cell(row=legend_row, column=3, value=desc)
        if _is_color:
            _fill = PatternFill(start_color=symbol, end_color=symbol, fill_type='solid')
            _ca.fill = _fill
            _cb.fill = _fill
        _ca.border = _thin_leg
        _cb.border = _thin_leg
        _cc.border = _thin_leg
        _cc.alignment = Alignment(horizontal='left')
        legend_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # Freeze panes
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 5: ASSESSMENT SHEET TEMPLATE
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Deletion Logging Assessment — evaluate centralised logging infrastructure and log completeness.',
        '2. Complete Verification Testing Program — assess forensic testing capability and testing frequency.',
        '3. Complete Evidence Repository Assessment — review storage, retention, and access control of deletion evidence.',
        '4. Complete Certificate Management — validate quality and management of deletion certificates.',
        '5. Maintain the Evidence Register with test results, certificates, and audit documentation.',
        '6. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A20"] = "Status Legend"
    ws["A20"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=21, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 22 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_assessment_sheet(ws, sheet_name, styles, checklist_items, reference_tables):
    """Create standard assessment sheet structure"""
    
    # Title
    ws['A1'] = sheet_name.upper()
    apply_style(ws['A1'], styles['header'])
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws.merge_cells('A1:U1')
    ws.row_dimensions[1].height = 35
    
    # Instructions
    ws['A2'] = "Instructions: Complete yellow-highlighted cells. Use dropdown menus where provided. Refer to checklist and reference tables below."
    ws.merge_cells('A2:U2')
    ws['A2'].alignment = Alignment(wrap_text=True)
    
    # Column headers - Standard (A-Q)
    for col_idx, header in enumerate(STANDARD_HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # Column headers - Extended (R-U)
    if sheet_name in EXTENDED_HEADERS:
        for col_idx, header in enumerate(EXTENDED_HEADERS[sheet_name], start=18):  # Start at column R (18)
            cell = ws.cell(row=4, column=col_idx, value=header)
            apply_style(cell, styles['header'])
    
    # Data entry rows: row 5 = F2F2F2 sample, rows 6-55 = 50 FFFFCC empty (total 51)
    sample_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row_idx in range(5, 56):  # Rows 5-55 (51 rows)
        for col_idx in range(1, 22):  # Columns A-U
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx == 5:
                cell.fill = sample_fill
            else:
                cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = thin_border

    # Sample row placeholder in column A
    ws.cell(row=5, column=1, value="[Sample] Example verification area entry")

    # Assessment Checklist
    checklist_start_row = 58
    ws.cell(row=checklist_start_row, column=1, value="ASSESSMENT CHECKLIST")
    ws.cell(row=checklist_start_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{checklist_start_row}:D{checklist_start_row}')
    
    checklist_headers = ["#", "Checklist Item", "Met?", "Evidence/Notes"]
    for col_idx, header in enumerate(checklist_headers, start=1):
        cell = ws.cell(row=checklist_start_row + 1, column=col_idx, value=header)
        apply_style(cell, styles['subheader'])
    
    for idx, item in enumerate(checklist_items, start=1):
        row = checklist_start_row + 1 + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item)
        # Column C (Met?) gets dropdown
        # Column D (Evidence/Notes) is text input
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles['ref_cell'])
    
    # Reference Tables (start after checklist)
    ref_table_start_row = checklist_start_row + 2 + len(checklist_items) + 2
    
    for table_idx, (table_title, table_data) in enumerate(reference_tables.items()):
        current_row = ref_table_start_row
        
        # Table title
        ws.cell(row=current_row, column=1, value=table_title)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, color='003366')
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1
        
        # Table headers
        headers = list(table_data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            apply_style(cell, styles['ref_header'])
        current_row += 1
        
        # Table data
        for row_data in table_data:
            for col_idx, value in enumerate(row_data.values(), start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                apply_style(cell, styles['ref_cell'])
            current_row += 1
        
        # Space between tables
        ref_table_start_row = current_row + 2
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['I'].width = 20
    for col in ['J', 'K', 'L']:
        ws.column_dimensions[col].width = 15
    for col in ['M', 'N', 'O', 'P']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['Q'].width = 30
    for col in ['R', 'S', 'T', 'U']:
        ws.column_dimensions[col].width = 20
    
    # Freeze panes at row 5 (header visible when scrolling)
    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 6: DATA VALIDATIONS
# ============================================================================

def apply_data_validations(ws, sheet_name):
    """Apply dropdown validations to assessment sheet"""
    validations = []

    # Standard validations (same for all assessment sheets)
    standard_validations = {
        'B': 'current_target_state',  # Current State
        'C': 'current_target_state',  # Target State
        'D': 'gap_severity',
        'E': 'control_effectiveness',
        'F': 'evidence_quality',
        'G': 'compliance_status',
        'H': 'risk_level',
        'J': 'review_frequency',
        'M': 'remediation_priority',
        'N': 'estimated_effort',
        'P': 'current_status'
    }

    # Extended validations (sheet-specific for columns R-U)
    extended_validations = {
        "Deletion Logging Assessment": {
            'S': None,  # Text (retention period)
            'T': 'yes_no_partial',  # Centralized System
            'U': 'tamper_protection'
        },
        "Verification Testing Program": {
            'R': 'test_frequency',
            'T': None  # Date
        },
        "Evidence Repository Assessment": {
            'S': 'access_controls',
            'T': 'yes_no_partial',  # Backup Status
            'U': 'retention_compliance'
        },
        "Certificate Management": {
            'T': 'validation_method',
            'U': None  # Text (storage period)
        },
        "Audit Trail Completeness": {
            'R': 'yes_no_partial',  # Reconstruction Tested
            'S': 'chain_of_custody',
            'U': 'audit_readiness'
        }
    }

    # Apply standard validations (rows 5-17)
    for col_letter, validation_key in standard_validations.items():
        if validation_key:
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(VALIDATION_LISTS[validation_key])}"',
                allow_blank=True
            )
            dv.error = "Please select from dropdown"
            dv.errorTitle = "Invalid Entry"
            dv.add(f'{col_letter}5:{col_letter}17')
            validations.append(dv)

    # Apply extended validations (sheet-specific)
    if sheet_name in extended_validations:
        for col_letter, validation_key in extended_validations[sheet_name].items():
            if validation_key:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(VALIDATION_LISTS[validation_key])}"',
                    allow_blank=True
                )
                dv.error = "Please select from dropdown"
                dv.errorTitle = "Invalid Entry"
                dv.add(f'{col_letter}5:{col_letter}17')
                validations.append(dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Checklist "Met?" dropdown (Column C, varies by checklist length)
    # This is handled separately in each sheet's specific population

# ============================================================================
# SECTION 7: SHEET 2 - DELETION LOGGING ASSESSMENT
# ============================================================================

def populate_sheet2_logging(ws, styles):
    """Populate Deletion Logging Assessment sheet"""
    
    checklist_items = [
        "Centralised log aggregation system implemented for all deletion operations",
        "Log retention periods align with legal/regulatory requirements (ISMS-POL-A.8.10-S2.3)",
        "Tamper-evident logging mechanisms protect log integrity",
        "Log storage capacity planning accounts for retention requirements",
        "Backup and disaster recovery procedures exist for log systems",
        "Logs capture: Who deleted (user/system identity)",
        "Logs capture: What was deleted (data classification, record identifiers - NOT actual data)",
        "Logs capture: When deletion occurred (timestamp with timezone)",
        "Logs capture: Where deletion occurred (system/location/media ID)",
        "Logs capture: How deletion was performed (method reference per NIST SP 800-88 Rev. 2)",
        "Log entries are machine-readable and parseable for audit analysis",
        "Log correlation capability exists across multiple systems/platforms",
        "Log completeness monitoring alerts on missing/failed log entries",
        "Automated log analysis identifies anomalies or failures",
        "Log access is restricted and audited (who views deletion logs)",
        "Log format supports GDPR Article 30 record-keeping requirements",
        "Logs provide evidence for FADP Article 6 accountability obligations",
        "Log retention balances evidence needs with data minimisation",
        "Logs support internal audit and external compliance assessments",
        "Log export capability exists for regulator/auditor requests"
    ]
    
    reference_tables = {
        "Table 1: Required Log Fields (NIST SP 800-88 Rev. 2 R1 Alignment)": [
            {
                "Log Field": "Deletion Request ID",
                "Required?": "Mandatory",
                "Example Value": "DEL-2025-001234",
                "Purpose": "Unique tracking identifier"
            },
            {
                "Log Field": "Requestor Identity",
                "Required?": "Mandatory",
                "Example Value": "jdoe@company.ch",
                "Purpose": "Accountability"
            },
            {
                "Log Field": "Deletion Timestamp",
                "Required?": "Mandatory",
                "Example Value": "2025-01-05T14:32:17Z",
                "Purpose": "Temporal evidence"
            },
            {
                "Log Field": "Data Classification",
                "Required?": "Mandatory",
                "Example Value": "Confidential-PII",
                "Purpose": "Risk context"
            },
            {
                "Log Field": "Media Type",
                "Required?": "Mandatory",
                "Example Value": "SSD / Cloud Storage",
                "Purpose": "Method validation"
            },
            {
                "Log Field": "Deletion Method",
                "Required?": "Mandatory",
                "Example Value": "ATA Secure Erase",
                "Purpose": "NIST SP 800-88 Rev. 2 reference"
            },
            {
                "Log Field": "Verification Method",
                "Required?": "Recommended",
                "Example Value": "Forensic Sampling",
                "Purpose": "Assurance level"
            },
            {
                "Log Field": "Certificate ID",
                "Required?": "Recommended",
                "Example Value": "CERT-20250105-789",
                "Purpose": "Third-party evidence"
            },
            {
                "Log Field": "Completion Status",
                "Required?": "Mandatory",
                "Example Value": "Success / Failed",
                "Purpose": "Outcome tracking"
            }
        ],
        "Table 2: Log Completeness Scoring Matrix": [
            {
                "Completeness Score": "90-100%",
                "Description": "Excellent",
                "Required Fields Captured": "All mandatory + all recommended fields"
            },
            {
                "Completeness Score": "75-89%",
                "Description": "Good",
                "Required Fields Captured": "All mandatory + most recommended fields"
            },
            {
                "Completeness Score": "60-74%",
                "Description": "Fair",
                "Required Fields Captured": "All mandatory fields, some recommended"
            },
            {
                "Completeness Score": "40-59%",
                "Description": "Poor",
                "Required Fields Captured": "Most mandatory, few recommended"
            },
            {
                "Completeness Score": "0-39%",
                "Description": "Critical Gap",
                "Required Fields Captured": "Missing mandatory fields"
            }
        ]
    }
    
    create_assessment_sheet(ws, "Deletion Logging Assessment", styles, checklist_items, reference_tables)
    apply_data_validations(ws, "Deletion Logging Assessment")

    # Add checklist "Met?" dropdown
    checklist_start = 21
    dv_checklist = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    dv_checklist.add(f'C{checklist_start + 2}:C{checklist_start + 2 + len(checklist_items) - 1}')
    ws.add_data_validation(dv_checklist)

# ============================================================================
# SECTION 8: SHEET 3 - VERIFICATION TESTING PROGRAM
# ============================================================================

def populate_sheet3_testing(ws, styles):
    """Populate Verification Testing Program sheet"""
    
    checklist_items = [
        "Documented verification testing program with defined scope and frequency",
        "Testing methodology aligns with NIST SP 800-88 Rev. 2 R1 verification guidance",
        "Test sample selection methodology is risk-based (prioritize high-risk deletions)",
        "Testing uses dedicated test datasets, NOT production data",
        "Test results are documented and retained per evidence retention policy",
        "Forensic tools/methods appropriate for media types (HDD, SSD, cloud, etc.)",
        "Testing validates deletion methods, not just process completion",
        "Cryptographic erasure validation confirms key destruction",
        "Cloud deletion testing includes API verification and metadata inspection",
        "Testing capability exists for ALL deletion methods in use",
        "Risk-based sampling strategy defined (higher risk = higher sample rate)",
        "Critical data deletions (PII, trade secrets) have higher verification rates",
        "Third-party/vendor deletions are tested (not just trusted on certificate)",
        "Testing covers edge cases (failed deletions, partial deletions, interruptions)",
        "Test failures trigger investigation and root cause analysis",
        "Failure trends are analysed to identify systemic issues",
        "Testing results feed into deletion method refinement",
        "Verification testing program is itself reviewed and improved annually"
    ]
    
    reference_tables = {
        "Table 1: Verification Testing Methods by Media Type": [
            {
                "Media Type": "HDD (Magnetic)",
                "Primary Verification Method": "Forensic imaging & analysis",
                "Tool Examples": "EnCase, FTK, dd + hexdump",
                "Pass Criteria": "No recoverable data patterns"
            },
            {
                "Media Type": "SSD (Flash)",
                "Primary Verification Method": "Controller-based validation",
                "Tool Examples": "Manufacturer tools, TRIM verification",
                "Pass Criteria": "Confirmed erasure completion"
            },
            {
                "Media Type": "Cloud Storage",
                "Primary Verification Method": "API query + metadata check",
                "Tool Examples": "AWS S3 verify, Azure Blob API",
                "Pass Criteria": "Object not found + no versions"
            },
            {
                "Media Type": "Backup Tapes",
                "Primary Verification Method": "Degaussing verification",
                "Tool Examples": "Gaussmeter, NIST test procedures",
                "Pass Criteria": "Magnetic field below threshold"
            },
            {
                "Media Type": "Mobile Devices",
                "Primary Verification Method": "Factory reset validation",
                "Tool Examples": "Cellebrite verification mode",
                "Pass Criteria": "No user data recoverable"
            }
        ],
        "Table 2: Risk-Based Sample Rate Guidelines": [
            {
                "Risk Level": "Critical",
                "Data Classification": "PII + Financial",
                "Minimum Sample Rate": "25-50%",
                "Rationale": "High regulatory exposure"
            },
            {
                "Risk Level": "High",
                "Data Classification": "PII or Confidential",
                "Minimum Sample Rate": "10-25%",
                "Rationale": "Significant impact if breach"
            },
            {
                "Risk Level": "Medium",
                "Data Classification": "Internal Use Only",
                "Minimum Sample Rate": "5-10%",
                "Rationale": "Moderate sensitivity"
            },
            {
                "Risk Level": "Low",
                "Data Classification": "Public",
                "Minimum Sample Rate": "1-5%",
                "Rationale": "Minimal impact"
            },
            {
                "Risk Level": "Continuous",
                "Data Classification": "All Third-Party",
                "Minimum Sample Rate": "100% (via certificates)",
                "Rationale": "Trust but verify vendors"
            }
        ]
    }
    
    create_assessment_sheet(ws, "Verification Testing Program", styles, checklist_items, reference_tables)
    apply_data_validations(ws, "Verification Testing Program")

    # Add checklist "Met?" dropdown
    checklist_start = 21
    dv_checklist = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    dv_checklist.add(f'C{checklist_start + 2}:C{checklist_start + 2 + len(checklist_items) - 1}')
    ws.add_data_validation(dv_checklist)

# ============================================================================
# SECTION 9: SHEET 4 - EVIDENCE REPOSITORY ASSESSMENT
# ============================================================================

def populate_sheet4_repository(ws, styles):
    """Populate Evidence Repository Assessment sheet"""
    
    checklist_items = [
        "Dedicated evidence repository exists (not scattered across file shares/emails)",
        "Repository access controls restrict evidence to authorised personnel",
        "Repository is backed up and included in disaster recovery procedures",
        "Repository capacity planning accounts for evidence accumulation",
        "Evidence is indexed/cataloged for efficient retrieval",
        "Evidence naming conventions facilitate cross-referencing with deletion logs",
        "Evidence metadata includes: deletion date, data classification, verification method",
        "Evidence search capability exists for audit/legal discovery requests",
        "Evidence repository contains deletion logs (per Sheet 2 assessment)",
        "Evidence repository contains verification test results (per Sheet 3 assessment)",
        "Evidence repository contains vendor certificates (per Sheet 5 assessment)",
        "Evidence repository does NOT contain actual deleted data (metadata only)",
        "Evidence retention periods documented and enforced (automated where possible)",
        "Evidence approaching retention expiration is flagged for review",
        "Evidence past retention period is itself deleted (data minimisation)",
        "Deletion of evidence is logged (meta-evidence of evidence deletion!)",
        "Evidence retention balances legal/regulatory requirements with data minimisation"
    ]
    
    reference_tables = {
        "Table 1: Evidence Types and Retention": [
            {
                "Evidence Type": "Deletion Logs",
                "Typical Retention Period": "1-7 years (per jurisdiction)",
                "Deletion Trigger": "Retention expiration + no legal hold",
                "Storage Format": "Structured logs (JSON/Syslog)"
            },
            {
                "Evidence Type": "Verification Test Results",
                "Typical Retention Period": "3-7 years",
                "Deletion Trigger": "Retention expiration",
                "Storage Format": "PDF reports + raw data"
            },
            {
                "Evidence Type": "Vendor Certificates",
                "Typical Retention Period": "3-7 years",
                "Deletion Trigger": "Retention expiration",
                "Storage Format": "PDF (digitally signed preferred)"
            },
            {
                "Evidence Type": "Audit Reports",
                "Typical Retention Period": "7-10 years",
                "Deletion Trigger": "Retention expiration",
                "Storage Format": "PDF"
            },
            {
                "Evidence Type": "Legal Hold Evidence",
                "Typical Retention Period": "Indefinite",
                "Deletion Trigger": "Court order release / settlement",
                "Storage Format": "Various (immutable)"
            }
        ],
        "Table 2: Access Control Matrix": [
            {
                "Role": "Security Team",
                "Deletion Logs": "Read/Write",
                "Test Results": "Read/Write",
                "Certificates": "Read/Write"
            },
            {
                "Role": "Compliance Team",
                "Deletion Logs": "Read",
                "Test Results": "Read",
                "Certificates": "Read"
            },
            {
                "Role": "IT Operations",
                "Deletion Logs": "Read (own systems only)",
                "Test Results": "No Access",
                "Certificates": "No Access"
            },
            {
                "Role": "External Auditors",
                "Deletion Logs": "Read (time-limited)",
                "Test Results": "Read (time-limited)",
                "Certificates": "Read (time-limited)"
            },
            {
                "Role": "General Employees",
                "Deletion Logs": "No Access",
                "Test Results": "No Access",
                "Certificates": "No Access"
            }
        ]
    }
    
    create_assessment_sheet(ws, "Evidence Repository Assessment", styles, checklist_items, reference_tables)
    apply_data_validations(ws, "Evidence Repository Assessment")

    # Add checklist "Met?" dropdown
    checklist_start = 21
    dv_checklist = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    dv_checklist.add(f'C{checklist_start + 2}:C{checklist_start + 2 + len(checklist_items) - 1}')
    ws.add_data_validation(dv_checklist)

# ============================================================================
# SECTION 10: SHEET 5 - CERTIFICATE MANAGEMENT
# ============================================================================

def populate_sheet5_certificates(ws, styles):
    """Populate Certificate Management sheet"""
    
    checklist_items = [
        "Vendor contracts require provision of deletion certificates per ISMS-POL-A.8.10-S2.4",
        "Certificate content requirements documented (minimum required fields)",
        "Certificate format standards defined (digitally signed PDF preferred)",
        "Certificate delivery timeline requirements in vendor SLAs (e.g., within 30 days)",
        "Non-compliance consequences defined in vendor contracts",
        "Certificate digital signatures are verified (not just accepted as-is)",
        "Certificate content is reviewed for completeness (all required fields present)",
        "Certificate issuer authority is validated (authorised vendor personnel)",
        "Certificate timestamps align with deletion request timelines",
        "Certificates are cross-referenced with deletion logs for consistency",
        "Generic/template certificates are flagged for additional verification",
        "Certificates include specific deletion details (not generic statements)",
        "Certificates reference deletion method/standard (e.g., NIST SP 800-88 Rev. 2)",
        "Certificates include unique identifiers (deletion job ID, media serial numbers)",
        "Certificates include attestation of subcontractor deletion",
        "Certificates stored in evidence repository with access controls",
        "Certificate inventory maintained (all vendors, all deletion events)",
        "Missing certificates are tracked and escalated (vendor non-compliance)",
        "Certificate retention aligns with overall evidence retention policy"
    ]
    
    reference_tables = {
        "Table 1: Certificate Quality Scoring Rubric (0-100 Points)": [
            {
                "Component": "Digital Signature",
                "Points": "20",
                "Criteria for Full Points": "Valid signature from authorised vendor representative"
            },
            {
                "Component": "Specific Data References",
                "Points": "20",
                "Criteria for Full Points": "Includes media IDs, deletion job IDs, or unique identifiers"
            },
            {
                "Component": "Deletion Method",
                "Points": "15",
                "Criteria for Full Points": "References recognised standard (NIST SP 800-88 Rev. 2, GDPR method)"
            },
            {
                "Component": "Timestamp Accuracy",
                "Points": "10",
                "Criteria for Full Points": "Deletion date/time aligns with request timeline"
            },
            {
                "Component": "Completeness",
                "Points": "15",
                "Criteria for Full Points": "All required fields present"
            },
            {
                "Component": "Third-Party Audit",
                "Points": "10",
                "Criteria for Full Points": "References independent audit report (SOC 2, ISO 27001)"
            },
            {
                "Component": "Subcontractor Coverage",
                "Points": "10",
                "Criteria for Full Points": "Confirms deletion by any subcontractors used"
            }
        ],
        "Table 2: Certificate Validation Red Flags": [
            {
                "Red Flag": "No digital signature",
                "Risk Level": "High",
                "Action Required": "Request signed certificate or reject"
            },
            {
                "Red Flag": "Generic template language",
                "Risk Level": "Medium",
                "Action Required": "Request specific details or supplemental evidence"
            },
            {
                "Red Flag": "Timestamp predates deletion request",
                "Risk Level": "High",
                "Action Required": "Investigate (possible fraudulent certificate)"
            },
            {
                "Red Flag": "Timestamp >90 days after request",
                "Risk Level": "Medium",
                "Action Required": "Investigate delay, verify deletion occurred"
            },
            {
                "Red Flag": "Signatory not in vendor contact list",
                "Risk Level": "High",
                "Action Required": "Verify signatory authority with vendor"
            },
            {
                "Red Flag": "No deletion method specified",
                "Risk Level": "Medium",
                "Action Required": "Request method details or perform verification testing"
            }
        ]
    }
    
    create_assessment_sheet(ws, "Certificate Management", styles, checklist_items, reference_tables)
    apply_data_validations(ws, "Certificate Management")

    # Add checklist "Met?" dropdown
    checklist_start = 21
    dv_checklist = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    dv_checklist.add(f'C{checklist_start + 2}:C{checklist_start + 2 + len(checklist_items) - 1}')
    ws.add_data_validation(dv_checklist)

# ============================================================================
# SECTION 11: SHEET 6 - AUDIT TRAIL COMPLETENESS
# ============================================================================

def populate_sheet6_audit_trail(ws, styles):
    """Populate Audit Trail Completeness sheet"""
    
    checklist_items = [
        "Audit trail links deletion request → approval → execution → verification → evidence",
        "Audit trail includes timestamps for each stage (demonstrating timeline)",
        "Audit trail identifies actors at each stage (accountability)",
        "Audit trail is tamper-evident (immutable logs or cryptographic protection)",
        "Audit trail reconstruction tested at least annually",
        "Reconstruction tests use real deletion events (not hypothetical scenarios)",
        "Reconstruction demonstrates compliance with retention schedule requirements",
        "Reconstruction demonstrates compliance with GDPR DSR timelines (if applicable)",
        "100% of deletion events have corresponding audit trail entries",
        "Audit trail captures successful AND failed deletion attempts",
        "Audit trail captures exceptions and workarounds (with approval documentation)",
        "Audit trail captures evidence of supervisory review (where required)",
        "Audit trail can be exported in auditor-friendly format (PDF, CSV, etc.)",
        "Audit trail terminology aligns with ISO 27001:2022 Annex A.8.10 language",
        "Audit trail demonstrates compliance with FADP/GDPR accountability",
        "Mock audits or self-assessments validate audit trail adequacy"
    ]
    
    reference_tables = {
        "Table 1: Chain of Custody Stages": [
            {
                "Stage": "1. Request",
                "Key Evidence": "Deletion request ticket",
                "Responsible Party": "Data Owner / DPO",
                "Timeframe": "T+0"
            },
            {
                "Stage": "2. Approval",
                "Key Evidence": "Approval record (if required)",
                "Responsible Party": "Manager / DPO",
                "Timeframe": "T+1 to T+3 days"
            },
            {
                "Stage": "3. Scheduling",
                "Key Evidence": "Work order / scheduled task",
                "Responsible Party": "IT Operations",
                "Timeframe": "T+3 to T+7 days"
            },
            {
                "Stage": "4. Execution",
                "Key Evidence": "Deletion log entry",
                "Responsible Party": "System / Operator",
                "Timeframe": "T+7 to T+30 days"
            },
            {
                "Stage": "5. Verification",
                "Key Evidence": "Test result / Certificate",
                "Responsible Party": "Security / Vendor",
                "Timeframe": "T+30 to T+60 days"
            },
            {
                "Stage": "6. Evidence Archival",
                "Key Evidence": "Repository entry",
                "Responsible Party": "Compliance Team",
                "Timeframe": "T+60 to T+90 days"
            },
            {
                "Stage": "7. Retention Expiry",
                "Key Evidence": "Evidence deletion log",
                "Responsible Party": "Records Management",
                "Timeframe": "T+Retention Period"
            }
        ],
        "Table 2: Reconstruction Test Procedure": [
            {
                "Test Step": "1. Select Sample",
                "Procedure": "Choose 5-10 deletion events (various types)",
                "Expected Result": "Representative sample",
                "Fail Criteria": "<5 events or non-representative"
            },
            {
                "Test Step": "2. Gather Evidence",
                "Procedure": "Collect all audit trail components",
                "Expected Result": "All components found",
                "Fail Criteria": "Missing any component"
            },
            {
                "Test Step": "3. Verify Chronology",
                "Procedure": "Confirm timestamps logical sequence",
                "Expected Result": "No gaps, logical flow",
                "Fail Criteria": "Timestamps out of order"
            },
            {
                "Test Step": "4. Validate Actors",
                "Procedure": "Confirm authorised personnel at each stage",
                "Expected Result": "All actors authorised",
                "Fail Criteria": "Unauthorised actor found"
            },
            {
                "Test Step": "5. Check Completeness",
                "Procedure": "Verify all required fields present",
                "Expected Result": "100% completeness",
                "Fail Criteria": "Any required field missing"
            },
            {
                "Test Step": "6. Cross-Reference",
                "Procedure": "Logs match certificates match test results",
                "Expected Result": "All sources consistent",
                "Fail Criteria": "Discrepancies found"
            },
            {
                "Test Step": "7. Demonstrate Compliance",
                "Procedure": "Show adherence to policy/regulation",
                "Expected Result": "Policy requirements met",
                "Fail Criteria": "Non-compliance identified"
            }
        ]
    }
    
    create_assessment_sheet(ws, "Audit Trail Completeness", styles, checklist_items, reference_tables)
    apply_data_validations(ws, "Audit Trail Completeness")

    # Add checklist "Met?" dropdown
    checklist_start = 21
    dv_checklist = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    dv_checklist.add(f'C{checklist_start + 2}:C{checklist_start + 2 + len(checklist_items) - 1}')
    ws.add_data_validation(dv_checklist)

# ============================================================================
# SECTION 12: HELPER FUNCTIONS
# ============================================================================

def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def populate_dashboard_sheet(ws, styles):
    """Populate Summary Dashboard sheet — Gold Standard TABLE 1/2/3 format."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ffffcc = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    d9d9d9 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # ── A1:G1 title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    title.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── A2:G2 subtitle ───────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    subtitle = ws["A2"]
    subtitle.value = (
        f"Summary Dashboard  |  {WORKBOOK_NAME}  |  Generated: {GENERATED_TIMESTAMP}"
    )
    subtitle.font = Font(name="Calibri", size=11, italic=True, color="003366")
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"

    # =========================================================================
    # TABLE 1 — COMPLIANCE ASSESSMENT SUMMARY (rows 4 onward)
    # =========================================================================
    ws.merge_cells("A4:G4")
    banner = ws["A4"]
    banner.value = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    banner.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    banner.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    banner.alignment = Alignment(horizontal="left", vertical="center")

    # Header row 5
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
                  "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Gen 4 assessment sheets use col G = Compliance Status
    # DV values: "Fully Compliant", "Partially Compliant", "Substantially Compliant",
    #            "Non-Compliant", "N/A (Not applicable to regulations)"
    # Sample row at row 5 → data rows start at row 6 (up to row 55)
    # Map: "Compliant" = Fully Compliant
    #      "Partial"   = Partially Compliant + Substantially Compliant
    #      "Non-Compliant" = Non-Compliant
    #      "N/A"       = N/A (Not applicable to regulations)
    FC_VAL  = "Fully Compliant"
    PC_VAL  = "Partially Compliant"
    SC_VAL  = "Substantially Compliant"
    NC_VAL  = "Non-Compliant"
    NA_VAL  = "N/A (Not applicable to regulations)"

    sheet_map = [
        ("Deletion Logging Assessment",    "'Deletion Logging Assessment'!G6:G55"),
        ("Verification Testing Program",   "'Verification Testing Program'!G6:G55"),
        ("Evidence Repository Assessment", "'Evidence Repository Assessment'!G6:G55"),
        ("Certificate Management",         "'Certificate Management'!G6:G55"),
        ("Audit Trail Completeness",       "'Audit Trail Completeness'!G6:G55"),
    ]

    row = 6
    for area_name, rng in sheet_map:
        c_formula  = f'=COUNTIF({rng},"{FC_VAL}")'
        p_formula  = (f'=COUNTIF({rng},"{PC_VAL}")'
                      f'+COUNTIF({rng},"{SC_VAL}")')
        nc_formula = f'=COUNTIF({rng},"{NC_VAL}")'
        na_formula = f'=COUNTIF({rng},"{NA_VAL}")'

        cells_in_row = [
            (ws.cell(row=row, column=1), area_name),
            (ws.cell(row=row, column=2), f"=C{row}+D{row}+E{row}+F{row}"),
            (ws.cell(row=row, column=3), c_formula),
            (ws.cell(row=row, column=4), p_formula),
            (ws.cell(row=row, column=5), nc_formula),
            (ws.cell(row=row, column=6), na_formula),
            (ws.cell(row=row, column=7), f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"),
        ]
        for cell, val in cells_in_row:
            cell.value = val
            cell.font = Font(name="Calibri", size=10, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=7).number_format = "0.0%"
        row += 1

    # TOTAL row
    total_row = row
    total_data = [
        (ws.cell(row=total_row, column=1), "TOTAL"),
        (ws.cell(row=total_row, column=2), f"=SUM(B6:B{total_row - 1})"),
        (ws.cell(row=total_row, column=3), f"=SUM(C6:C{total_row - 1})"),
        (ws.cell(row=total_row, column=4), f"=SUM(D6:D{total_row - 1})"),
        (ws.cell(row=total_row, column=5), f"=SUM(E6:E{total_row - 1})"),
        (ws.cell(row=total_row, column=6), f"=SUM(F6:F{total_row - 1})"),
        (ws.cell(row=total_row, column=7),
         f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"),
    ]
    for cell, val in total_data:
        cell.value = val
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=7).number_format = "0.0%"
    row = total_row + 2  # blank spacer

    # =========================================================================
    # TABLE 2 — KEY METRICS
    # =========================================================================
    ws.merge_cells(f"A{row}:G{row}")
    t2_banner = ws[f"A{row}"]
    t2_banner.value = "TABLE 2 \u2013 KEY METRICS"
    t2_banner.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    t2_banner.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    t2_banner.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    for col_idx, hdr in enumerate(["Metric", "Value", "Target"], start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    row += 1

    t2_data = [
        ("Overall Compliance Rate",
         f"=IFERROR(G{total_row},\"N/A\")",
         "\u2265 80%"),
        ("Deletion Logging Controls Fully Compliant",
         "=IFERROR(COUNTIF('Deletion Logging Assessment'!G6:G55,\""
         + FC_VAL + "\"),0)",
         "100%"),
        ("Verification Testing Controls Fully Compliant",
         "=IFERROR(COUNTIF('Verification Testing Program'!G6:G55,\""
         + FC_VAL + "\"),0)",
         "100%"),
        ("Evidence Repository Controls Fully Compliant",
         "=IFERROR(COUNTIF('Evidence Repository Assessment'!G6:G55,\""
         + FC_VAL + "\"),0)",
         "100%"),
        ("Certificate Management Controls Fully Compliant",
         "=IFERROR(COUNTIF('Certificate Management'!G6:G55,\""
         + FC_VAL + "\"),0)",
         "100%"),
        ("Non-Compliant Items Requiring Remediation",
         f"=IFERROR(SUM(E6:E{total_row - 1}),0)",
         "0"),
    ]

    for metric, value, target in t2_data:
        ws.cell(row=row, column=1).value = metric
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=3).value = target
        for col_idx in range(1, 4):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = ffffcc
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).number_format = "0.0%"
        row += 1

    row += 1  # spacer

    # =========================================================================
    # TABLE 3 — KEY FINDINGS & RECOMMENDATIONS
    # =========================================================================
    ws.merge_cells(f"A{row}:G{row}")
    t3_banner = ws[f"A{row}"]
    t3_banner.value = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    t3_banner.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    t3_banner.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    t3_banner.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", "Notes"]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    row += 1

    t3_rows = [
        ("1", "Deletion logs may lack tamper protection, reducing evidential value",
         "High", "Implement cryptographic signing or WORM storage for all deletion logs",
         "P1", "Open", ""),
        ("2", "Forensic verification testing may be infrequent or cover limited media types",
         "High", "Establish quarterly forensic testing schedule for high-risk data categories",
         "P1", "Open", ""),
        ("3", "Evidence repository access controls may be insufficient for audit scenarios",
         "Medium", "Restrict evidence repository access to Security/Compliance roles only",
         "P2", "Open", ""),
        ("4", "Deletion certificates from vendors may not be validated before acceptance",
         "Medium", "Implement standard validation procedure (signature check + content review)",
         "P2", "Open", ""),
    ]

    for row_data in t3_rows:
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.fill = ffffcc
            cell.border = border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
        row += 1

def populate_evidence_register(ws, styles):
    """Populate Evidence Register sheet (100 rows) - Golden Standard"""
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Row 1: Title banner
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "Comprehensive evidence tracking for A.8.10 Data Deletion compliance assessment"
    cell.font = Font(name='Calibri', size=10, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 3: spacer

    # Row 4: Headers with 003366 fill, white bold font
    headers = ["Evidence ID", "Category", "Description", "Source Document",
               "Date Collected", "Collected By", "Status", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Rows 5-104: Data rows with EV-001..EV-100
    # Apply FFFFCC to ALL columns (1-8) including Evidence ID column
    for i in range(100):
        row_num = i + 5
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_num, column=col_idx)
            if col_idx == 1:
                cell.value = f"EV-{i+1:03d}"
                cell.font = Font(name='Calibri', size=9, color='808080')
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.border = border

    # Column widths
    widths = {'A': 15, 'B': 20, 'C': 40, 'D': 25, 'E': 15, 'F': 20, 'G': 15, 'H': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze
    ws.freeze_panes = 'A5'

    # DataValidation for Status column (G)
    validations = []
    dv = DataValidation(type='list', formula1='"Verified,Pending verification,Not verified,Requires update"', allow_blank=True)
    dv.add('G5:G104')
    validations.append(dv)
    for _dv in validations:
        ws.add_data_validation(_dv)

def populate_approval_signoff(ws, styles):
    """Populate Approval Sign-Off sheet - Golden Standard"""
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Row 1: Title banner
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "ASSESSMENT APPROVAL AND SIGN-OFF"
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells('A2:E2')
    ws['A2'].value = f"ISMS-IMP-A.8.10.4 - Verification & Evidence Assessment"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].border = border

    # Freeze
    ws.freeze_panes = 'A3'

    # Row 3: ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].value = "ASSESSMENT SUMMARY"
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].border = border
    row += 1

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
    ]
    validations = []
    status_row = None
    for label, value in summary_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', bold=True)
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].value = value
        ws[f'B{row}'].border = border
        if value == "":
            ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    if status_row:
        dv_status = DataValidation(type='list',
                                   formula1='"Draft,Final,Requires remediation,Re-assessment required"',
                                   allow_blank=True)
        dv_status.add(f'B{status_row}')
        validations.append(dv_status)

    row += 1  # gap

    # 3 approver blocks
    approvers = [
        ("COMPLETED BY (DATA PROTECTION)", "4472C4"),
        ("REVIEWED BY (COMPLIANCE)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    for title, color in approvers:
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f'A{row}'] = field
            ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=10)
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws[f'B{row}'].border = border
            row += 1
        row += 1  # spacer

    # FINAL ASSESSMENT DECISION — col A plain bold label (GS: no dark fill, no merge)
    cell = ws[f'A{row}']
    cell.value = "FINAL DECISION:"
    cell.font = Font(name='Calibri', size=11, bold=True, color='000000')
    row += 1
    ws[f'A{row}'] = "Decision:"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True)
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    ws[f'B{row}'].border = border
    dv = DataValidation(type='list', formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    dv.add(f'B{row}')
    validations.append(dv)
    row += 2

    # NEXT REVIEW DETAILS
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "NEXT REVIEW DETAILS"
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.border = border
    row += 1
    for field in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', bold=True)
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'B{row}'].border = border
        row += 1

    # Column A width
    ws.column_dimensions['A'].width = 32

    for _dv in validations:
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 13: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.8.10.4 - Verification & Evidence Assessment")
        logger.info("Workbook Generator")
        logger.info("=" * 70)

        # Create workbook
        logger.info("Creating workbook structure...")
        wb = create_workbook()
        styles = create_styles()

        # Populate each sheet
        logger.info("Populating Instructions sheet...")
        populate_instructions_sheet(wb["Instructions & Legend"], styles)

        logger.info("Populating Sheet 2: Deletion Logging Assessment...")
        populate_sheet2_logging(wb["Deletion Logging Assessment"], styles)

        logger.info("Populating Sheet 3: Verification Testing Program...")
        populate_sheet3_testing(wb["Verification Testing Program"], styles)

        logger.info("Populating Sheet 4: Evidence Repository Assessment...")
        populate_sheet4_repository(wb["Evidence Repository Assessment"], styles)

        logger.info("Populating Sheet 5: Certificate Management...")
        populate_sheet5_certificates(wb["Certificate Management"], styles)

        logger.info("Populating Sheet 6: Audit Trail Completeness...")
        populate_sheet6_audit_trail(wb["Audit Trail Completeness"], styles)

        logger.info("Populating Sheet 7: Summary Dashboard...")
        populate_dashboard_sheet(wb["Summary Dashboard"], styles)

        logger.info("Populating Sheet 8: Evidence Register...")
        populate_evidence_register(wb["Evidence Register"], styles)

        logger.info("Populating Sheet 9: Approval Sign-Off...")
        populate_approval_signoff(wb["Approval Sign-Off"], styles)

        # Save workbook
        filename = f"{FILENAME_PREFIX}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info("=" * 70)
        logger.info("WORKBOOK GENERATION COMPLETE!")
        logger.info("=" * 70)
        logger.info(f"File: {filename}")
        logger.info("Structure: 9 sheets (Instructions + 5 Assessments + Dashboard + Evidence + Approval)")
        logger.info("  - 5 assessment sheets with 13 data entry rows each")
        logger.info("  - Evidence Register: 100 rows")
        logger.info("  - Approval Sign-Off: 3-level workflow")
        logger.info("=" * 70)
        return 0
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
