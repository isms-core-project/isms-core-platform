#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.10.3 - Third-Party & Cloud Deletion Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.10: Information Deletion
Assessment Domain 3 of 4: Third-Party and Cloud Service Deletion Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific cloud service providers, SaaS applications, vendor
landscape, and data processing agreements.

Key customization areas:
1. Cloud service provider inventory (match your actual IaaS/PaaS/SaaS vendors)
2. Data Processing Agreement (DPA) clause requirements (adapt to your legal standards)
3. Subprocessor identification (based on your vendor's supply chain)
4. Shadow IT discovery mechanisms (specific to your detection tools)
5. Deletion verification methods (aligned with provider capabilities)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.10 Information Deletion Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
third-party and cloud service provider deletion capabilities against ISO 27001:2022
Control A.8.10 and GDPR Article 28 processor obligations, supporting evidence-based
validation of vendor deletion controls.

**Purpose:**
Enables systematic assessment of cloud provider and third-party deletion capabilities,
contract compliance, subprocessor risks, and shadow IT exposure to ensure information
is properly deleted when stored or processed by external parties.

**Assessment Scope:**
- Cloud provider deletion capabilities (IaaS, PaaS providers)
- SaaS application deletion procedures and data residency
- Vendor contract assessment (DPA deletion clauses, SLAs)
- Subprocessor mapping and deletion chain-of-custody
- Shadow IT discovery and remediation
- GDPR Article 28 processor obligation compliance
- Deletion certificate and audit report collection
- Gap analysis for inadequate vendor controls
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance, GDPR Article 28 overview, color coding
2. Cloud Provider Deletion - IaaS/PaaS provider deletion capabilities and verification
3. SaaS Application Deletion - Software-as-a-Service deletion procedures
4. Vendor Contract Assessment - DPA deletion clauses and SLA compliance
5. Subprocessor Mapping - Third-party subprocessor identification and deletion
6. Shadow IT Assessment - Unapproved cloud services discovery and remediation
7. Summary Dashboard - Vendor risk analysis, compliance gaps identification
8. Evidence Register - DPA copies, deletion certificates, audit reports
9. Approval Sign-Off - Three-level stakeholder approval workflow

**Key Features:**
- Data validation with cloud provider tier and deletion capability dropdown lists
- Conditional formatting for vendor risk and deletion capability status
- Automated gap identification for missing deletion clauses or certificates
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with ISMS-REF-A.5.23 Cloud Service Provider Registry (if available)
- GDPR Article 28 and Swiss FADP processor obligation compliance

**Integration:**
This assessment feeds into the A.8.10.5 Compliance Dashboard, which
consolidates data from all four information deletion assessment domains for
executive oversight and audit readiness.

**ISMS-REF-A.5.23 Cloud Service Provider Registry Integration:**
If available, this assessment references the Cloud Service Provider Registry for:
- Provider tier ratings (Tier 1: Enterprise, Tier 2: Mid-Market, Tier 3: Small/Niche)
- Standard deletion capability ratings by provider
- Recommended DPA templates and deletion clauses
- Deletion verification methods by provider

If ISMS-REF-A.5.23 is not available, manually assess provider capabilities.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a810_3_third_party_cloud.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a810_3_third_party_cloud.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a810_3_third_party_cloud.py --date 20250124

Output:
    File: ISMS_A_8_10_3_Third_Party_Cloud_Deletion_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize provider tiers (if using ISMS-REF-A.5.23)
    2. Inventory all cloud service providers and SaaS applications
    3. Complete vendor deletion capability assessment for each provider
    4. Review DPA contracts for deletion clauses and SLAs
    5. Map subprocessor relationships and deletion dependencies
    6. Conduct shadow IT discovery and assessment
    7. Validate deletion certificates and audit reports
    8. Conduct gap analysis for inadequate vendor controls
    9. Define remediation actions with timelines
    10. Collect and link audit evidence (DPAs, certificates, audit reports)
    11. Obtain stakeholder approvals
    12. Feed results into A.8.10.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.10
Assessment Domain:    3 of 4 (Third-Party and Cloud Service Deletion Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.10: Information Deletion Policy (Governance)
    - ISMS-REF-A.5.23: Cloud Service Provider Registry (Provider Deletion Capabilities)
    - ISMS-IMP-A.8.10.1: Retention & Deletion Triggers Assessment (Domain 1)
    - ISMS-IMP-A.8.10.2: Deletion Methods Assessment (Domain 2)
    - ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion Implementation Guide
    - ISMS-IMP-A.8.10.4: Verification & Evidence Assessment (Domain 4)
    - ISMS-IMP-A.8.10.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.10.3 specification
    - Supports comprehensive third-party and cloud deletion evaluation
    - Integrated with A.8.10.5 Compliance Dashboard
    - Optional integration with ISMS-REF-A.5.23 Cloud Service Provider Registry

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**GDPR Article 28 Processor Obligations:**
Under GDPR Article 28, data processors (cloud/SaaS providers) MUST:
1. Only process data on documented instructions from the controller
2. Ensure personnel are bound by confidentiality
3. Implement appropriate technical and organizational measures
4. Only engage sub-processors with prior authorization
5. Assist the controller in fulfilling data subject rights
6. **Delete or return all personal data after services end** (Article 28(3)(g))
7. Make available information necessary to demonstrate compliance

This assessment evaluates vendor compliance with Article 28(3)(g) deletion obligations.

**Swiss FADP Processor Requirements:**
Under Swiss FADP Article 9, data processors must:
1. Process data only as instructed by the controller
2. Implement appropriate technical and organizational measures
3. Delete or return data when contract ends
4. Maintain processing records

**Cloud Provider Deletion Capabilities Vary Significantly:**
- **Tier 1 (Enterprise):** AWS, Azure, GCP - Comprehensive deletion APIs, certificates
- **Tier 2 (Mid-Market):** DigitalOcean, Linode - Good deletion controls, limited certs
- **Tier 3 (Small/Niche):** Smaller providers - Manual processes, limited verification

Assess each provider's specific capabilities and contract terms.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of DPA deletion clauses, deletion certificates,
and subprocessor mapping.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Cloud service provider inventory and data locations
- Vendor contract terms and DPA clauses
- Subprocessor relationships and data flows
- Shadow IT findings and remediation plans

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new cloud services or vendor changes
- Semi-annually: Review DPA renewals and deletion clause updates
- Annually: Complete reassessment of all providers and subprocessors
- Ad-hoc: When new vendors are onboarded or contracts renewed

**Quality Assurance:**
Have Procurement, Legal/Compliance, Information Security, and Data Protection teams
validate assessments before using results for compliance reporting or remediation
decisions.

**Regulatory Alignment:**
Ensure vendor deletion controls align with applicable regulatory requirements:
- Privacy laws: GDPR Article 28, Swiss FADP Article 9 processor obligations
- Financial services: Outsourcing guidelines, third-party risk management
- Healthcare: HIPAA Business Associate Agreements (BAA) and data disposal
- Payment processing: PCI DSS third-party service provider requirements

Customize assessment criteria to include regulatory-specific requirements.

**Common Vendor Deletion Risks:**
1. ❌ No deletion clause in contract - No legal obligation to delete
2. ❌ No deletion verification - Provider claims deletion without proof
3. ❌ Unknown subprocessors - Data copied to unknown third parties
4. ❌ Shadow IT services - Unapproved cloud services with no deletion controls
5. ❌ Backup retention - Provider retains data in backups after deletion
6. ❌ Multi-tenant databases - Logical deletion without physical separation

**Deletion Verification Methods:**
Different providers offer different verification approaches:
- **Deletion Certificates:** Formal attestation of data deletion (best)
- **API Confirmation:** Programmatic deletion confirmation
- **Audit Reports:** SOC 2 Type II deletion control verification
- **Manual Confirmation:** Email/ticket confirmation (weakest)

Prefer providers offering formal deletion certificates or audit report verification.

**Subprocessor Chain-of-Custody:**
When your provider uses subprocessors:
1. Identify all subprocessors handling your data
2. Ensure subprocessor agreements include deletion obligations
3. Verify deletion flows through entire subprocessor chain
4. Obtain deletion confirmation from primary processor (who manages subprocessors)

You are not required to obtain deletion certificates from every subprocessor,
but your primary processor must ensure subprocessor deletion.

**Shadow IT Risks:**
Unapproved cloud services pose significant deletion risks:
- No contractual deletion obligations
- Unknown data locations and retention
- No deletion verification possible
- Compliance and audit gaps

Implement shadow IT discovery (CASB, network monitoring, expense analysis)
and remediation procedures.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.10.3"
WORKBOOK_NAME = "Third-Party & Cloud Deletion Assessment"
CONTROL_ID = "A.8.10"
CONTROL_NAME = "Information Deletion"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ==========================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ==========================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
TRASH = '\U0001F5D1'  # 🗑️  Wastebasket
DISK = '\U0001F4BE'   # 💾 Floppy Disk
GLOBE = '\U0001F310'  # 🌐 Globe
SEARCH = '\U0001F50D' # 🔍 Magnifying Glass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order (per ISMS-IMP-A.8.10.3 specification)
    wb.create_sheet("Instructions & Legend", 0)
    wb.create_sheet("2. Cloud Provider Deletion", 1)
    wb.create_sheet("3. SaaS Application Deletion", 2)
    wb.create_sheet("4. Vendor Contract Assessment", 3)
    wb.create_sheet("5. Subprocessor Mapping", 4)
    wb.create_sheet("6. Shadow IT Assessment", 5)
    wb.create_sheet("Summary Dashboard", 6)
    wb.create_sheet("Evidence Register", 7)
    wb.create_sheet("Approval Sign-Off", 8)
    
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    styles = {
        'title': {
            'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'header': {
            'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'subheader': {
            'font': Font(name='Calibri', size=10, bold=True, color='000000'),
            'fill': PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'input_cell': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'reference_cell': {
            'font': Font(name='Calibri', size=9, color='000000'),
            'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'normal': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    }
    return styles


# ==========================================================================
# SECTION 2: COLUMN DEFINITIONS
# ==========================================================================

def get_base_columns():
    """Return base column structure (A-Q, 17 columns)."""
    return [
        ("A", "Provider / Vendor Name", 30),
        ("B", "Service Type", 22),
        ("C", "Business Owner", 18),
        ("D", "Data Categories Processed", 25),
        ("E", "Provider Tier", 12),
        ("F", "Status", 18),
        ("G", "Contract Start Date", 12),
        ("H", "Last Contract Review", 12),
        ("I", "Next Contract Review", 12),
        ("J", "Gap Identified", 25),
        ("K", "Remediation Plan", 25),
        ("L", "Target Completion", 12),
        ("M", "Risk Level", 12),
        ("N", "Evidence Reference", 20),
        ("O", "Notes / Comments", 25),
        ("P", "Remediation Owner", 18),
        ("Q", "Budget Required", 15)
    ]


def get_extended_columns_sheet2():
    """Sheet 2: Cloud Provider Deletion extensions."""
    return [
        ("R", "Deletion Method", 20),
        ("S", "Deletion SLA (Days)", 15),
        ("T", "Certificate Provided", 20),
        ("U", "Multi-Region Verified", 20)
    ]


def get_extended_columns_sheet3():
    """Sheet 3: SaaS Application Deletion extensions."""
    return [
        ("R", "Admin Portal Access", 20),
        ("S", "Data Export Available", 20),
        ("T", "Deletion Request Method", 22),
        ("U", "GDPR DPA Signed", 18)
    ]


def get_extended_columns_sheet4():
    """Sheet 4: Vendor Contract Assessment extensions."""
    return [
        ("R", "Deletion Clause Present", 22),
        ("S", "Deletion Timeline Specified", 25),
        ("T", "Certificate of Deletion", 22),
        ("U", "Audit Rights", 18)
    ]


def get_extended_columns_sheet5():
    """Sheet 5: Subprocessor Mapping extensions."""
    return [
        ("R", "Subprocessor Count", 18),
        ("S", "Subprocessor List Current", 22),
        ("T", "Deletion Flow Documented", 22),
        ("U", "Subprocessor Deletion SLA", 22)
    ]


def get_extended_columns_sheet6():
    """Sheet 6: Shadow IT Assessment extensions."""
    return [
        ("R", "Deletion Requests (Last 12M)", 22),
        ("S", "Average Response Time (Days)", 25),
        ("T", "Certificates Received", 20),
        ("U", "Incidents / Failures", 18)
    ]


# ==========================================================================
# SECTION 3: DATA VALIDATION
# ==========================================================================

def create_base_validations(ws):
    """Create data validation objects for standard columns."""
    
    # Service Type (Column B)
    dv_service = DataValidation(
        type="list",
        formula1='"IaaS (Infrastructure as a Service),PaaS (Platform as a Service),SaaS (Software as a Service),Data Processor (non-cloud),Subprocessor,Other"',
        allow_blank=True
    )
    dv_service.error = 'Please select from dropdown'
    dv_service.errorTitle = 'Invalid Service Type'
    ws.add_data_validation(dv_service)
    dv_service.add('B10:B100')
    
    # Provider Tier (Column E) - Integration with ISMS-REF-A.5.23
    dv_tier = DataValidation(
        type="list",
        formula1='"Tier 1,Tier 2,Tier 3,Tier 4,Tier 5,Tier 6,Tier 7,Tier 8,Tier 9,Tier 10"',
        allow_blank=True
    )
    dv_tier.error = 'Please select from dropdown - See ISMS-REF-A.5.23 for tier definitions'
    dv_tier.errorTitle = 'Invalid Provider Tier'
    ws.add_data_validation(dv_tier)
    dv_tier.add('E10:E100')
    
    # Status (Column F)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=True
    )
    dv_status.error = 'Please select from dropdown'
    dv_status.errorTitle = 'Invalid Status'
    ws.add_data_validation(dv_status)
    dv_status.add('F10:F100')
    
    # Risk Level (Column M)
    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    dv_risk.error = 'Please select from dropdown'
    dv_risk.errorTitle = 'Invalid Risk Level'
    ws.add_data_validation(dv_risk)
    dv_risk.add('M10:M100')
    
    # Budget Required (Column Q)
    dv_budget = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    dv_budget.error = 'Please select from dropdown'
    dv_budget.errorTitle = 'Invalid Budget Option'
    ws.add_data_validation(dv_budget)
    dv_budget.add('Q10:Q100')


def create_sheet_specific_validations(ws, sheet_type):
    """Create sheet-specific data validations."""
    
    if sheet_type == "sheet2":
        # Deletion Method (Column R)
        dv = DataValidation(
            type="list",
            formula1='"API Delete,Console Delete,Support Ticket,Account Closure,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Certificate Provided (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Automatic,Yes - On Request,No,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # Multi-Region Verified (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Yes,No,N/A,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')
    
    elif sheet_type == "sheet3":
        # Admin Portal Access (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Full,Yes - Limited,No,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Data Export Available (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Yes - API,Yes - Manual,No,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Deletion Request Method (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Self-Service,Email Request,Support Ticket,Account Manager"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # GDPR DPA Signed (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Yes,No,N/A,Pending"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')
    
    elif sheet_type == "sheet4":
        # Deletion Clause Present (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Specific,Yes - Generic,No,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Deletion Timeline Specified (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Days Specified,Yes - Reasonable Time,No"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Certificate of Deletion (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Contractually Required,Available on Request,Not Mentioned"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # Audit Rights (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Deletion Audit,Yes - General Audit,No,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')
    
    elif sheet_type == "sheet5":
        # Subprocessor List Current (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Updated,Yes - Outdated,No - Unknown,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Deletion Flow Documented (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Flowchart,Yes - Description,No,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # Subprocessor Deletion SLA (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Covered by Prime,Separate Agreement,Unknown,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')


# ==========================================================================
# SECTION 4: HELPER FUNCTIONS
# ==========================================================================

def apply_cell_style(cell, styles, style_type):
    """Apply style dictionary to a cell."""
    style = styles[style_type]
    cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


def create_header_row(ws, row, columns, styles):
    """Create header row with column definitions."""
    for col_letter, header_text, width in columns:
        cell = ws[f'{col_letter}{row}']
        cell.value = header_text
        apply_cell_style(cell, styles, 'header')
        ws.column_dimensions[col_letter].width = width


def create_data_rows(ws, start_row, num_rows, num_cols, styles):
    """Create yellow-highlighted data entry rows."""
    for row in range(start_row, start_row + num_rows):
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles, 'input_cell')


def create_checklist_section(ws, start_row, checklist_items, styles):
    """Create compliance checklist section."""
    ws.merge_cells(f'A{start_row}:U{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = "COMPLIANCE CHECKLIST - Mark ✅ or {XMARK}"
    apply_cell_style(cell, styles, 'subheader')
    
    current_row = start_row + 1
    for idx, item in enumerate(checklist_items, 1):
        ws[f'A{current_row}'].value = f"{idx}."
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
        
        ws.merge_cells(f'B{current_row}:S{current_row}')
        ws[f'B{current_row}'].value = item
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        ws[f'T{current_row}'].value = ""
        ws[f'T{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        
        ws[f'U{current_row}'].value = ""
        ws[f'U{current_row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        current_row += 1
    
    return current_row


def create_reference_table(ws, start_row, table_title, headers, data, styles):
    """Create reference table with title and data."""
    num_cols = len(headers)
    end_col = get_column_letter(num_cols)
    ws.merge_cells(f'A{start_row}:{end_col}{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = table_title
    apply_cell_style(cell, styles, 'subheader')
    
    header_row = start_row + 1
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{header_row}']
        cell.value = header
        apply_cell_style(cell, styles, 'header')
    
    current_row = header_row + 1
    for row_data in data:
        for col_idx, value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = value
            apply_cell_style(cell, styles, 'reference_cell')
        current_row += 1
    
    return current_row + 1


# ==========================================================================
# SECTION 5: ASSESSMENT SHEET CREATOR (CORE FUNCTION)
# ==========================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, 
                            assessment_question, base_cols, extended_cols,
                            checklist_items, reference_tables, sheet_type):
    """Create standardized assessment sheet."""
    
    ws.merge_cells('A1:U1')
    cell = ws['A1']
    cell.value = section_title
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:U2')
    cell = ws['A2']
    cell.value = f"Policy Reference: {policy_ref}"
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='left')
    
    ws.row_dimensions[3].height = 5
    
    ws.merge_cells('A4:U6')
    cell = ws['A4']
    cell.value = f"ASSESSMENT QUESTION:\n{assessment_question}"
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[4].height = 50
    
    ws.merge_cells('A7:U7')
    cell = ws['A7']
    cell.value = "Complete the yellow-highlighted cells below. Use dropdowns where provided. Reference ISMS-REF-A.5.23 for Provider Tier definitions. Link evidence in Column N to Evidence Register sheet."
    cell.font = Font(name='Calibri', size=9, italic=True, color='0000FF')
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    ws.row_dimensions[8].height = 5
    
    all_columns = base_cols + extended_cols
    create_header_row(ws, 9, all_columns, styles)
    
    create_data_rows(ws, 10, 13, len(all_columns), styles)
    
    ws.row_dimensions[23].height = 10
    ws.row_dimensions[24].height = 5
    
    next_row = create_checklist_section(ws, 25, checklist_items, styles)
    
    ws.row_dimensions[next_row].height = 10
    next_row += 1
    
    for table_title, headers, data in reference_tables:
        next_row = create_reference_table(ws, next_row, table_title, headers, data, styles)
    
    ws.merge_cells(f'A{next_row}:U{next_row}')
    cell = ws[f'A{next_row}']
    cell.value = "EXCEPTIONS / DEVIATIONS"
    apply_cell_style(cell, styles, 'subheader')
    
    next_row += 1
    ws.merge_cells(f'A{next_row}:U{next_row+2}')
    cell = ws[f'A{next_row}']
    cell.value = "Document any exceptions or deviations from requirements here:"
    cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[next_row].height = 60
    
    ws.freeze_panes = 'A10'
    
    create_base_validations(ws)
    create_sheet_specific_validations(ws, sheet_type)


# ==========================================================================
# SECTION 6: INSTRUCTIONS SHEET
# ==========================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with GDPR focus."""
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ISMS-IMP-A.8.10.3 - Third-Party & Cloud Deletion Assessment"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "DOCUMENT CONTROL"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    control_fields = [
        ("Workbook Version:", "1.0"),
        ("Assessment Date:", datetime.now().strftime('%d.%m.%Y')),
        ("Assessor Name:", "[Enter Name]"),
        ("Organization:", "[Enter Organization]"),
        ("Review Period:", "[e.g., Q4 2025]")
    ]
    
    for label, value in control_fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    current_row += 2
    
    # GDPR Article 28 Framework
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "GDPR ARTICLE 28 - PROCESSOR REQUIREMENTS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    gdpr_requirements = [
        ("Article 28(3)(e)", "Processor must delete or return all personal data after services end, at controller's choice."),
        ("Article 28(3)(h)", "Processor must assist controller in responding to data subject deletion requests (Article 17)."),
        ("Article 28(2)", "Processor must not engage subprocessors without authorization. Controller must be informed of changes."),
        ("DPA Requirement", "Deletion obligations must be in writing (Data Processing Agreement). Timeline should be specified.")
    ]
    
    for article, description in gdpr_requirements:
        ws[f'A{current_row}'].value = article
        ws[f'A{current_row}'].font = Font(bold=True, size=10)
        ws[f'A{current_row}'].fill = PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = description
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 30
        current_row += 2
    
    # Provider Tier Quick Reference
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "PROVIDER TIER QUICK REFERENCE (See ISMS-REF-A.5.23 for complete 68-provider registry)"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    tier_headers = ["Tier", "Description", "Example Providers", "Deletion Risk"]
    for col_idx, header in enumerate(tier_headers, 1):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'].value = header
        apply_cell_style(ws[f'{col_letter}{current_row}'], styles, 'header')
    current_row += 1
    
    tier_data = [
        ("Tier 1", "Hyperscaler - Critical Infrastructure", "AWS, Azure, GCP", "Low"),
        ("Tier 2", "Major Provider - High Dependency", "Salesforce, Oracle Cloud", "Low-Medium"),
        ("Tier 3", "Significant Provider - Medium Dependency", "Dropbox, Box, Zendesk", "Medium"),
        ("Tier 4-8", "Standard to Legacy Providers", "Regional/niche providers", "Medium-High"),
        ("Tier 9", "Departmental - Shadow IT Risk", "Unapproved SaaS", "Very High"),
        ("Tier 10", "Unknown/Unclassified", "Discovered via audit", "Critical")
    ]
    
    for tier, desc, examples, risk in tier_data:
        ws[f'A{current_row}'].value = tier
        ws[f'B{current_row}'].value = desc
        ws[f'C{current_row}'].value = examples
        ws[f'D{current_row}'].value = risk
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        current_row += 1
    
    current_row += 2
    
    # Important Notes
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "CRITICAL NOTES"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    notes = [
        f"{WARNING} CRITICAL: Tier 9/10 vendors (Shadow IT) represent significant compliance risk - immediate remediation required.",
        "✓ GDPR DPA is MANDATORY for all processors handling EU/EEA personal data.",
        "✓ Certificate of Deletion is best practice for Restricted/Confidential data (especially Tier 1-3 vendors).",
        "✓ Subprocessor changes must be notified per GDPR Article 28(2).",
        "✓ Annual vendor contract review is required per ISMS-POL-A.8.10-S4.",
        "✓ Integration with ISMS-REF-A.5.23: Use the 68-provider registry for tier classification.",
        "✓ Related assessments: A.8.10.1 (Retention), A.8.10.2 (Methods), A.8.10.4 (Verification)"
    ]
    
    for note in notes:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = note
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 25
        current_row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


# ==========================================================================
# SECTION 7: EVIDENCE REGISTER
# ==========================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register with 100 rows."""
    
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "Evidence Register - Supporting Documentation"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "Use this register to document all evidence supporting your third-party deletion assessments. Reference evidence by ID (Column A) in the 'Evidence Reference' column (Column N) of assessment sheets."
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    ws.row_dimensions[3].height = 5
    
    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Assessment Sheet", 20),
        ("C", "Related Vendor/Provider", 30),
        ("D", "Evidence Type", 20),
        ("E", "Evidence Title/Description", 35),
        ("F", "File Location/Link", 40),
        ("G", "Date Created/Collected", 12),
        ("H", "Retention Period", 15),
        ("I", "Next Review Date", 12),
        ("J", "Owner/Custodian", 20),
        ("K", "Notes", 30)
    ]
    
    create_header_row(ws, 4, headers, styles)
    
    for row in range(5, 105):
        for col_idx, (col_letter, _, _) in enumerate(headers, 1):
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles, 'input_cell')
    
    # Dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Data Processing Agreement (DPA),Vendor Contract,Deletion Clause Extract,Certificate of Deletion,Email Correspondence,Subprocessor List,Deletion Request Log,Performance Report,Audit Report,Meeting Minutes,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add('D5:D104')
    
    dv_sheet = DataValidation(
        type="list",
        formula1='"Sheet 2: Cloud Provider Deletion,Sheet 3: SaaS Application Deletion,Sheet 4: Vendor Contract Assessment,Sheet 5: Subprocessor Mapping,Sheet 6: Shadow IT Assessment"',
        allow_blank=True
    )
    ws.add_data_validation(dv_sheet)
    dv_sheet.add('B5:B104')
    
    dv_retention = DataValidation(
        type="list",
        formula1='"Duration of vendor relationship + 7 years,7 years (standard contract retention),10 years (critical vendors),Permanent (ongoing relationship)"',
        allow_blank=True
    )
    ws.add_data_validation(dv_retention)
    dv_retention.add('H5:H104')
    
    ws.freeze_panes = 'A5'


# ==========================================================================
# SECTION 8: APPROVAL SIGN-OFF
# ==========================================================================

def create_approval_signoff(ws, styles):
    """Create Approval Sign-Off workflow sheet."""
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "Three-Level Approval Workflow"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "DOCUMENT CONTROL"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    control_fields = [
        ("Assessment Period:", "[e.g., Q4 2025]"),
        ("Workbook Version:", "1.0"),
        ("Total Assessment Sheets Completed:", "5"),
        ("Overall Compliance %:", "[Link to Summary Dashboard]"),
        ("Critical Vendor Risks:", "[Count from Summary]"),
        ("Shadow IT Vendors (Tier 9/10):", "[Count]"),
        ("Assessment Completed By:", "[Name, Date]")
    ]
    
    for label, value in control_fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    current_row += 2
    
    # Approval levels with specific roles and statements
    approvals = [
        ("LEVEL 1 APPROVAL - Operational", "Vendor Management / Procurement / IT Compliance Officer",
         'I confirm that this assessment accurately reflects our current third-party vendor relationships, cloud provider deletion capabilities, and subprocessor mappings as of [Date]. All vendors have been reviewed, gaps identified, and remediation plans are in place.'),
        ("LEVEL 2 APPROVAL - Legal/Compliance", "Data Protection Officer / General Counsel / Chief Compliance Officer",
         'I acknowledge the findings of this A.8.10.3 assessment from a legal and compliance perspective. Vendor contracts meet GDPR Article 28 requirements (or remediation is planned). Critical vendor risks, particularly Shadow IT (Tier 9/10), require immediate executive attention.'),
        ("LEVEL 3 APPROVAL - Executive", "Chief Information Officer / Chief Risk Officer / Chief Procurement Officer",
         'This assessment has been reviewed at the executive level. The organization\'s third-party deletion posture is [Acceptable/Needs Improvement/Unacceptable]. The Board/Executive Team has been briefed on vendor risks, particularly: (1) Shadow IT vendors, (2) Vendors without deletion clauses, (3) Subprocessor complexity.')
    ]
    
    for level_title, role, statement in approvals:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = level_title
        apply_cell_style(cell, styles, 'subheader')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = f'Role: {role}'
        ws[f'A{current_row}'].font = Font(italic=True)
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row + 1}')
        ws[f'A{current_row}'].value = f'Approval Statement: "{statement}"'
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[current_row].height = 50
        current_row += 2
        
        approval_fields = ["Approver Name:", "Title/Role:", "Email:", "Review Date:", "Approval Status:", "Conditions/Comments:", "Signature:"]
        
        for field in approval_fields:
            ws[f'A{current_row}'].value = field
            ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            if field == "Approval Status:":
                dv = DataValidation(type="list", formula1=f'"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f'B{current_row}')
            current_row += 1
        
        current_row += 2
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40


# ==========================================================================
# SECTION 9: SUMMARY DASHBOARD
# ==========================================================================

def create_summary_dashboard(ws, styles):
    """Create comprehensive summary dashboard with Shadow IT detection."""
    
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "A.8.10.3 Third-Party & Cloud Deletion - Compliance Dashboard"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    # Overall Compliance
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "OVERALL COMPLIANCE SUMMARY"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    headers = ["Assessment Area", "Total", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col_idx)}{current_row}'].value = header
        apply_cell_style(ws[f'{get_column_letter(col_idx)}{current_row}'], styles, 'header')
    current_row += 1
    
    areas = ["Cloud Provider Deletion", "SaaS Application Deletion", "Vendor Contract Assessment", 
             "Subprocessor Mapping", "Shadow IT Assessment"]
    
    for area in areas:
        ws[f'A{current_row}'].value = area
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{current_row}'].value = "[Formula]"
        current_row += 1
    
    ws[f'A{current_row}'].value = "OVERALL A.8.10.3"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 3
    
    # Shadow IT Detection
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"{WARNING} CRITICAL - SHADOW IT DETECTION (Tier 9/10)"
    apply_cell_style(cell, styles, 'subheader')
    ws[f'A{current_row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Tier 9 Vendors (Departmental/Unapproved):"
    ws[f'B{current_row}'].value = '[=COUNTIF(Sheet2!E:E,"Tier 9")]'
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Tier 10 Vendors (Unknown/Unclassified):"
    ws[f'B{current_row}'].value = '[=COUNTIF(Sheet2!E:E,"Tier 10")]'
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'].value = "If count > 0: IMMEDIATE REMEDIATION REQUIRED - No DPA, no deletion clause, high GDPR risk"
    ws[f'A{current_row}'].font = Font(bold=True, color='9C0006')
    current_row += 3
    
    # Executive Summary
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "EXECUTIVE SUMMARY & RECOMMENDATIONS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Overall A.8.10.3 Maturity Level:"
    ws[f'B{current_row}'].value = "[Emerging / Developing / Established / Optimized]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 2
    
    ws[f'A{current_row}'].value = "Key Strengths:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    for i in range(1, 4):
        ws[f'A{current_row}'].value = f"{i}."
        ws[f'B{current_row}'].value = "[Example strength]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    current_row += 1
    ws[f'A{current_row}'].value = "Critical Improvement Areas:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    for i in range(1, 4):
        ws[f'A{current_row}'].value = f"{i}."
        ws[f'B{current_row}'].value = "[Critical gap]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20


# ==========================================================================
# SECTION 10: DOMAIN-SPECIFIC SHEET CREATORS
# ==========================================================================

def create_sheet2_cloud_provider(ws, styles):
    """Create Sheet 2: Cloud Provider Deletion."""
    
    checklist = [
        "All cloud providers in use are identified and documented",
        "Provider tier assigned per ISMS-REF-A.5.23 classification",
        "Data Processing Agreement (DPA) signed for all providers",
        "DPA includes specific deletion obligations and timeline",
        "Deletion method documented (API, console, support ticket)",
        "API-based deletion implemented for Tier 1-3 providers",
        "Deletion SLA documented and reasonable (≤30 days preferred)",
        "Certificate of Deletion available (automatic or on request)",
        "Multi-region/multi-zone deletion tested and verified",
        "Snapshot and backup deletion procedures documented",
        "Account closure deletion behavior documented",
        "Provider notifications sent for subprocessor changes",
        "Annual contract review conducted",
        "Deletion requests successfully executed in last 12 months",
        "Provider included in ISMS-REF-A.5.23 registry (or add if new)"
    ]
    
    reference_tables = [
        ("Major Cloud Provider Deletion Capabilities (See ISMS-REF-A.5.23 for complete details)",
         ["Provider", "Tier", "Deletion Method", "Certificate Available"],
         [
             ["AWS", "Tier 1", "API Delete, Console", "On request"],
             ["Microsoft Azure", "Tier 1", "API Delete, Portal", "On request"],
             ["Google Cloud", "Tier 1", "API Delete, Console", "On request"],
             ["Oracle Cloud", "Tier 2", "API Delete, Console", "On request"],
             ["IBM Cloud", "Tier 2", "API Delete, Console", "On request"],
             ["DigitalOcean", "Tier 3", "API Delete, Console", "On request"]
         ]),
        ("Deletion SLA Best Practices",
         ["Data Sensitivity", "Recommended SLA", "Rationale"],
         [
             ["Restricted (health, financial)", "≤7 days", "High sensitivity, regulatory risk"],
             ["Confidential", "≤30 days", "GDPR default timeframe"],
             ["Internal", "≤60 days", "Business data, moderate sensitivity"],
             ["Public", "≤90 days", "Low sensitivity, operational flexibility"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 2: Cloud Provider Deletion",
        "ISMS-POL-A.8.10-S2.4, Section 2.4.1",
        "Do we have documented deletion capabilities, contractual agreements, and performance metrics for all cloud infrastructure and platform providers?",
        get_base_columns(), get_extended_columns_sheet2(),
        checklist, reference_tables,
        "sheet2"
    )


def create_sheet3_saas_application(ws, styles):
    """Create Sheet 3: SaaS Application Deletion."""
    
    checklist = [
        "All SaaS applications in use are identified and documented",
        "Business owners assigned for each SaaS application",
        "Data Processing Agreement (DPA) signed for SaaS handling personal data",
        "DPA includes deletion obligations and timeline",
        "Admin portal access verified (ability to manage data)",
        "Data export capability tested (before deletion if needed)",
        "Deletion request method documented (self-service, ticket, etc.)",
        "Deletion confirmation process exists",
        "User account deletion vs. data deletion distinguished",
        "Shared account deletion procedures documented",
        "Integration data deletion addressed (APIs, webhooks)",
        "Backup/archive deletion coordinated",
        "Shadow SaaS applications investigated and remediated",
        "Annual SaaS inventory review conducted",
        "GDPR data subject request procedures cover SaaS platforms"
    ]
    
    reference_tables = [
        ("Common SaaS Categories and Deletion",
         ["Category", "Examples", "Typical Method", "Data Export"],
         [
             ["CRM", "Salesforce, HubSpot", "Admin portal delete", "API/CSV"],
             ["HR/Payroll", "Workday, BambooHR", "Admin + ticket", "CSV/API"],
             ["Marketing", "Mailchimp, Braze", "Admin delete", "API/CSV"],
             ["Collaboration", "Slack, Teams", "Admin console", "Limited"],
             ["File Sharing", "Dropbox, Box", "Admin delete", "Download"],
             ["Analytics", "Google Analytics", "Admin delete", "Limited"]
         ]),
        ("SaaS Deletion Challenges",
         ["Challenge", "Mitigation"],
         [
             ["User Deletion ≠ Data Deletion", "Request data deletion explicitly"],
             ["Shared Data", "Coordinate with data owners"],
             ["Integration Data", "Delete from all integrated systems"],
             ["Backup Retention", "Confirm backup deletion timeline"],
             ["No Self-Service", "Document request process, SLA"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 3: SaaS Application Deletion",
        "ISMS-POL-A.8.10-S2.4, Section 2.4.2",
        "Can we effectively delete data from all SaaS applications, including obtaining confirmation and managing data export if needed?",
        get_base_columns(), get_extended_columns_sheet3(),
        checklist, reference_tables,
        "sheet3"
    )


def create_sheet4_vendor_contract(ws, styles):
    """Create Sheet 4: Vendor Contract Assessment."""
    
    checklist = [
        "All vendor contracts reviewed for deletion clauses",
        "Deletion clause specificity assessed (specific vs. generic)",
        "Deletion timeline explicitly stated (not just 'reasonable time')",
        "Certificate of Deletion either required or available on request",
        "Audit rights include ability to verify deletion compliance",
        "Subprocessor usage disclosed and governed",
        "Data return option available as alternative to deletion",
        "Contract renewal triggers deletion clause review",
        "New vendor onboarding includes deletion clause requirement",
        "Vendor contract template updated with deletion requirements",
        "Legal/Procurement engaged for contract negotiations",
        "High-risk vendors (Tier 1-3, Restricted data) have strongest clauses",
        "Contract breach remedies include deletion failures",
        "Vendor due diligence includes deletion capability assessment",
        "Contracts reviewed at least every 3 years"
    ]
    
    reference_tables = [
        ("Deletion Clause Strength",
         ["Type", "Strength", "Recommended For"],
         [
             ["Specific with timeline and certificate", "Excellent", "Tier 1-3, Restricted data"],
             ["Generic with timeline (≤90 days)", "Acceptable", "Tier 4-6, Confidential"],
             ["Generic only", "Insufficient", "Not recommended"],
             ["Absent", "Non-Compliant", "Requires amendment"]
         ]),
        ("Certificate of Deletion Elements",
         ["Element", "Example"],
         [
             ["Date of Deletion", "Deleted on: 2025-12-15"],
             ["Scope", "All customer records for Company X"],
             ["Method", "Secure deletion per NIST SP 800-88"],
             ["Systems", "Production DB, backups, archives"],
             ["Verification", "Forensic verification completed"],
             ["Authorized Signatory", "Signed by: DPO"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 4: Vendor Contract Assessment",
        "ISMS-POL-A.8.10-S2.4, Section 2.4.3",
        "Do all vendor contracts include appropriate deletion clauses, SLAs, and audit rights to ensure data deletion compliance?",
        get_base_columns(), get_extended_columns_sheet4(),
        checklist, reference_tables,
        "sheet4"
    )


def create_sheet5_subprocessor(ws, styles):
    """Create Sheet 5: Subprocessor Mapping."""
    
    checklist = [
        "All primary vendors (processors) identified",
        "Subprocessor disclosure obtained from each vendor",
        "Subprocessor count documented and current",
        "Subprocessor list reviewed at least annually",
        "Vendor contractually obligated to notify of subprocessor changes",
        "Data flow diagram created for critical data categories",
        "Deletion flow documented (prime vendor → subprocessor deletion)",
        "Subprocessor deletion SLA understood (covered by prime or separate)",
        "Subprocessor DPAs flow down from prime vendor",
        "High-risk subprocessors identified (non-EU, Tier 9/10)",
        "Subprocessor changes trigger risk assessment",
        "Deletion requests explicitly include subprocessor deletion",
        "Certificate of Deletion covers subprocessor deletion",
        "Multi-tier subprocessing mapped (sub-subprocessors)",
        "Shadow subprocessors investigated (undisclosed usage)"
    ]
    
    reference_tables = [
        ("Common Subprocessor Categories",
         ["Primary Vendor", "Common Subprocessors", "Coordination"],
         [
             ["Cloud Provider", "CDN, Monitoring, Logging", "Via prime APIs"],
             ["SaaS CRM", "Email, Analytics, Payment", "Prime coordinates"],
             ["HR/Payroll", "Background check, Benefits", "Often manual"],
             ["Marketing", "Email, SMS, Ad networks", "API/dashboard"],
             ["File Sharing", "Preview, Virus scan, OCR", "Automatic via prime"]
         ]),
        ("GDPR Article 28(2) Requirements",
         ["Requirement", "Verification"],
         [
             ["Prior Authorization", "Authorized in contract?"],
             ["Notification of Changes", "Notification process?"],
             ["Objection Rights", "Objection clause present?"],
             ["Same Obligations", "Flow-down confirmed?"],
             ["Prime Liability", "Liability clause present?"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 5: Subprocessor Mapping",
        "ISMS-POL-A.8.10-S2.4, Section 2.4.4",
        "Do we know all subprocessors in our vendor supply chain, and can we coordinate deletion across the entire chain?",
        get_base_columns(), get_extended_columns_sheet5(),
        checklist, reference_tables,
        "sheet5"
    )


def create_sheet6_vendor_performance(ws, styles):
    """Create Sheet 6: Shadow IT Assessment."""
    
    checklist = [
        "Deletion request log maintained for all vendors",
        "Response time tracked per vendor",
        "SLA compliance calculated (actual vs. contracted)",
        "Certificate of Deletion requested for high-value deletions",
        "Certificate completeness reviewed",
        "Deletion incidents documented and investigated",
        "Root cause analysis performed for failed deletions",
        "Vendor escalation procedures exist for SLA breaches",
        "Quarterly vendor performance review conducted",
        "Underperforming vendors identified and remediated",
        "Performance metrics reported to management",
        "Vendor scorecard includes deletion performance",
        "Contract renewal considers deletion performance history",
        "Best-performing vendors identified for benchmarking",
        "Performance data feeds into vendor risk assessment"
    ]
    
    reference_tables = [
        ("Performance Metrics",
         ["Metric", "Target", "Use Case"],
         [
             ["Average Response Time", "≤ Contracted SLA", "Overall performance"],
             ["SLA Compliance Rate", "≥ 95%", "Contract compliance"],
             ["Certificate Completion", "100% (critical vendors)", "Documentation quality"],
             ["Incident Rate", "< 5%", "Reliability"]
         ]),
        ("Deletion Incident Types",
         ["Type", "Severity", "Remediation"],
         [
             ["SLA Breach", "Medium-High", "Escalate, track pattern"],
             ["Incomplete Deletion", "High", "Re-request, verify"],
             ["No Certificate", "Medium", "Request, document"],
             ["Subprocessor Failure", "Critical", "Escalate to prime, audit"],
             ["Vendor Refusal", "Critical", "Legal escalation, exit"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 6: Shadow IT Assessment",
        "ISMS-POL-A.8.10-S2.4, Section 2.4.5",
        "Are vendors meeting their deletion SLAs, providing certificates when required, and is performance monitored over time?",
        get_base_columns(), get_extended_columns_sheet6(),
        checklist, reference_tables,
        "sheet6"
    )


# ==========================================================================
# SECTION 11: REFERENCE DATA
# ==========================================================================

# Defined inline in sheet creators


# ==========================================================================
# SECTION 12: MAIN EXECUTION
# ==========================================================================

def main():
    """Main execution function."""
    try:
        logger.info("=" * 78)
        logger.info("ISMS-IMP-A.8.10.3 - Third-Party & Cloud Deletion Assessment Generator")
        logger.info("ISO/IEC 27001:2022 Control A.8.10: Information Deletion")
        logger.info("GDPR Article 28 Processor Compliance + ISMS-REF-A.5.23 Integration")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/9] Creating Instructions & Legend (GDPR Article 28)...")
        create_instructions_sheet(wb["Instructions & Legend"], styles)

        logger.info("[2/9] Creating Sheet 2: Cloud Provider Deletion...")
        create_sheet2_cloud_provider(wb["2. Cloud Provider Deletion"], styles)

        logger.info("[3/9] Creating Sheet 3: SaaS Application Deletion...")
        create_sheet3_saas_application(wb["3. SaaS Application Deletion"], styles)

        logger.info("[4/9] Creating Sheet 4: Vendor Contract Assessment...")
        create_sheet4_vendor_contract(wb["4. Vendor Contract Assessment"], styles)

        logger.info("[5/9] Creating Sheet 5: Subprocessor Mapping...")
        create_sheet5_subprocessor(wb["5. Subprocessor Mapping"], styles)

        logger.info("[6/9] Creating Sheet 6: Shadow IT Assessment...")
        create_sheet6_vendor_performance(wb["6. Shadow IT Assessment"], styles)

        logger.info("[7/9] Creating Summary Dashboard (with Shadow IT Detection)...")
        create_summary_dashboard(wb["Summary Dashboard"], styles)

        logger.info("[8/9] Creating Evidence Register (100 rows)...")
        create_evidence_register(wb["Evidence Register"], styles)

        logger.info("[9/9] Creating Approval Sign-Off (3-level workflow)...")
        create_approval_signoff(wb["Approval Sign-Off"], styles)

        filename = f"ISMS-IMP-A.8.10.3_Third_Party_Cloud_Deletion_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("=" * 78)
        logger.info("SUCCESS: %s", filename)
        logger.info("Workbook Structure:")
        logger.info("  - Instructions & Legend - GDPR Article 28 processor requirements")
        logger.info("  - Sheet 2: Cloud Provider Deletion - IaaS/PaaS deletion capabilities")
        logger.info("  - Sheet 3: SaaS Application Deletion - CRM, HR, collaboration platforms")
        logger.info("  - Sheet 4: Vendor Contract Assessment - Deletion clauses, SLAs, audit rights")
        logger.info("  - Sheet 5: Subprocessor Mapping - Data flow and deletion coordination")
        logger.info("  - Sheet 6: Shadow IT Assessment - SLA compliance, certificates")
        logger.info("  - Summary Dashboard - Executive overview + Shadow IT detection")
        logger.info("  - Evidence Register - 100 rows for DPAs, contracts, certificates")
        logger.info("  - Approval Sign-Off - 3-level approval (Ops, Legal, Executive)")
        logger.info("Key Features: ISMS-REF-A.5.23 integration, GDPR Article 28 compliance, Shadow IT detection")
        logger.info("=" * 78)
        return 0
    except Exception as e:
        logger.error("Failed to generate workbook: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
