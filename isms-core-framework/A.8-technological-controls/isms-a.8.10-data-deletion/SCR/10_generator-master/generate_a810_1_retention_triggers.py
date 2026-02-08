#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.10.1 - Retention & Deletion Triggers Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.10: Information Deletion
Assessment Domain 1 of 4: Retention Schedule & Deletion Trigger Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data categories, retention requirements, regulatory
obligations, and deletion trigger mechanisms.

Key customization areas:
1. Data category taxonomy (match your organization's data classification scheme)
2. Retention periods (adapt to your regulatory and business requirements)
3. Deletion trigger types (automated vs. manual trigger mechanisms)
4. Legal hold procedures (specific to your legal/compliance processes)
5. Data subject rights workflows (GDPR/FADP compliance procedures)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.10 Information Deletion Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
retention schedules and deletion trigger mechanisms against ISO 27001:2022 Control
A.8.10 requirements, supporting evidence-based validation of systematic information
deletion practices across the data lifecycle.

**Purpose:**
Enables systematic assessment of retention schedule definition, documentation,
and enforcement mechanisms to ensure information is deleted when no longer required
for business, legal, or regulatory purposes.

**Assessment Scope:**
- Complete data category inventory and classification
- Retention period definition and regulatory alignment
- Automated and manual deletion trigger configuration
- Legal hold management procedures and tracking
- Data subject rights (GDPR Article 17 / Swiss FADP erasure requests)
- Retention schedule documentation and approval
- Deletion trigger testing and verification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance, regulatory context, color coding
2. Data Category Registry - Complete inventory of all data types and classifications
3. Retention Schedule Compliance - Legal/regulatory alignment for retention periods
4. Deletion Trigger Configuration - Automated and manual deletion mechanisms
5. Legal Hold Management - Procedures to suspend deletion when legally required
6. Data Subject Rights - GDPR/FADP erasure request handling workflows
7. Summary Dashboard - Compliance overview, KPIs, critical gaps identification
8. Evidence Register - Audit evidence tracking and documentation linkage
9. Approval Sign-Off - Three-level stakeholder approval workflow

**Key Features:**
- Data validation with data classification and retention period dropdown lists
- Conditional formatting for retention compliance status indication
- Automated gap identification for missing retention schedules
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with GDPR Article 17 and Swiss FADP requirements
- Legal hold tracking and suspension procedures

**Integration:**
This assessment feeds into the A.8.10.5 Compliance Dashboard, which
consolidates data from all four information deletion assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a810_1_retention_triggers.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a810_1_retention_triggers.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a810_1_retention_triggers.py --date 20250124

Output:
    File: ISMS_A_8_10_1_Retention_Deletion_Triggers_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize retention periods to match your regulatory context
    2. Inventory all data categories across all systems and processes
    3. Complete retention schedule for each data category
    4. Validate deletion triggers are properly configured
    5. Review legal hold procedures and tracking mechanisms
    6. Document data subject rights handling workflows
    7. Conduct gap analysis for missing or incomplete retention schedules
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (retention policies, deletion logs)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.10.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.10
Assessment Domain:    1 of 4 (Retention Schedule & Deletion Trigger Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.10: Information Deletion Policy (Governance)
    - ISMS-IMP-A.8.10.1: Retention & Deletion Triggers Implementation Guide
    - ISMS-IMP-A.8.10.2: Deletion Methods Assessment (Domain 2)
    - ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion Assessment (Domain 3)
    - ISMS-IMP-A.8.10.4: Verification & Evidence Assessment (Domain 4)
    - ISMS-IMP-A.8.10.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.10.1 specification
    - Supports comprehensive retention schedule and deletion trigger evaluation
    - Integrated with A.8.10.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Regulatory Requirements:**
Retention periods must comply with applicable regulations including:
- GDPR: Article 5(1)(e) storage limitation principle
- Swiss FADP: Article 6(3) data processing principles (proportionality)
- Industry-specific: PCI DSS (payment data), HIPAA (health records), etc.
- Jurisdiction-specific: Local data protection and archival laws

This assessment provides a framework - customize retention periods based on
your organization's legal, regulatory, and business requirements.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of retention schedule documentation, deletion
trigger configuration, and legal hold procedures.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Data category inventory and classification
- Retention periods and business justifications
- System inventory and data locations
- Legal hold information

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new data categories or regulatory changes
- Semi-annually: Update retention periods for legislative changes
- Annually: Complete reassessment of all data categories and deletion triggers
- Ad-hoc: When new systems are deployed or regulatory requirements change

**Quality Assurance:**
Have Data Protection Officers, Legal/Compliance teams, and Business Unit Owners
validate assessments before using results for compliance reporting or remediation
decisions.

**Regulatory Alignment:**
Ensure retention periods align with applicable regulatory requirements:
- Privacy laws: GDPR Article 17, Swiss FADP erasure requirements
- Financial services: Basel III, MiFID II, Dodd-Frank retention requirements
- Healthcare: HIPAA minimum necessary and retention requirements
- Payment processing: PCI DSS data retention and disposal requirements
- Public sector: Government archival and public records laws

Customize assessment criteria to include regulatory-specific requirements.

**Data Subject Rights (GDPR Article 17 / Swiss FADP):**
Organizations must be able to:
1. Identify all systems containing a data subject's information
2. Validate the identity of the requestor
3. Assess whether deletion exceptions apply (legal obligations, public interest)
4. Execute deletion across all systems within required timeframes
5. Document what was deleted and what was retained (with justification)
6. Notify third parties who received the data
7. Respond to the data subject with confirmation

This assessment evaluates the organization's capability to fulfill these
obligations systematically and within regulatory timeframes.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.10.1"
WORKBOOK_NAME = "Retention & Deletion Triggers Assessment"
CONTROL_ID = "A.8.10"
CONTROL_NAME = "Information Deletion"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ==========================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ==========================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
TRASH = '\U0001F5D1'  # 🗑️  Wastebasket
DISK = '\U0001F4BE'   # 💾 Floppy Disk
GLOBE = '\U0001F310'  # 🌐 Globe
SEARCH = '\U0001F50D' # 🔍 Magnifying Glass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    wb.create_sheet("Instructions & Legend", 0)
    wb.create_sheet("2. Data Category Registry", 1)
    wb.create_sheet("3. Retention Schedule Compliance", 2)
    wb.create_sheet("4. Deletion Trigger Configuration", 3)
    wb.create_sheet("5. Legal Hold Management", 4)
    wb.create_sheet("6. Data Subject Requests", 5)
    wb.create_sheet("Summary Dashboard", 6)
    wb.create_sheet("Evidence Register", 7)
    wb.create_sheet("Approval Sign-Off", 8)
    
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    styles = {
        'title': {
            'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'header': {
            'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'subheader': {
            'font': Font(name='Calibri', size=10, bold=True, color='000000'),
            'fill': PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'input_cell': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_compliant': {
            'font': Font(name='Calibri', size=10, bold=True, color='006100'),
            'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_partial': {
            'font': Font(name='Calibri', size=10, bold=True, color='9C5700'),
            'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_noncompliant': {
            'font': Font(name='Calibri', size=10, bold=True, color='9C0006'),
            'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'reference_cell': {
            'font': Font(name='Calibri', size=9, color='000000'),
            'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'normal': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    }
    return styles


# ==========================================================================
# SECTION 2: COLUMN DEFINITIONS
# ==========================================================================

def get_base_columns():
    """Return base column structure (A-Q, 17 columns)."""
    return [
        ("A", "Data Category / System Name", 30),
        ("B", "Data Classification", 22),
        ("C", "Business Owner", 18),
        ("D", "Retention Period", 20),
        ("E", "Legal/Regulatory Basis", 20),
        ("F", "Status", 18),
        ("G", "Implementation Date", 12),
        ("H", "Last Review Date", 12),
        ("I", "Next Review Date", 12),
        ("J", "Gap Identified", 25),
        ("K", "Remediation Plan", 25),
        ("L", "Target Completion", 12),
        ("M", "Risk Level", 12),
        ("N", "Evidence Reference", 20),
        ("O", "Notes / Comments", 25),
        ("P", "Remediation Owner", 18),
        ("Q", "Budget Required", 15)
    ]


def get_extended_columns_sheet2():
    """Sheet 2: Data Category Registry extensions."""
    return [
        ("R", "Primary Storage Location", 22),
        ("S", "Volume/Records", 18),
        ("T", "Contains PII/SPI", 18)
    ]


def get_extended_columns_sheet3():
    """Sheet 3: Retention Schedule Compliance extensions."""
    return [
        ("R", "Retention Calculation Method", 22),
        ("S", "Event Trigger Description", 25),
        ("T", "Backup Retention Aligned", 18)
    ]


def get_extended_columns_sheet4():
    """Sheet 4: Deletion Trigger Configuration extensions."""
    return [
        ("R", "Trigger Type", 18),
        ("S", "Trigger Frequency", 18),
        ("T", "Legal Hold Check Integrated", 22)
    ]


def get_extended_columns_sheet5():
    """Sheet 5: Legal Hold Management extensions."""
    return [
        ("R", "Active Legal Holds", 15),
        ("S", "Legal Hold Notification Process", 25),
        ("T", "Hold Review Frequency", 18)
    ]


def get_extended_columns_sheet6():
    """Sheet 6: Data Subject Requests extensions."""
    return [
        ("R", "Average Response Time (Days)", 20),
        ("S", "GDPR/FADP Applicable", 18),
        ("T", "Request Volume (Last 12 Months)", 22)
    ]


# ==========================================================================
# SECTION 3: DATA VALIDATION
# ==========================================================================

def create_base_validations(ws):
    """Create data validation objects for standard columns."""
    
    # Data Classification (Column B)
    dv_classification = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=True
    )
    dv_classification.error = 'Please select from dropdown'
    dv_classification.errorTitle = 'Invalid Classification'
    ws.add_data_validation(dv_classification)
    dv_classification.add(f'B10:B100')
    
    # Retention Period (Column D)
    dv_retention = DataValidation(
        type="list",
        formula1='"30 days,60 days,90 days,6 months,1 year,2 years,3 years,5 years,7 years,10 years,Permanent,Until Event Occurs,Other (specify in notes)"',
        allow_blank=True
    )
    dv_retention.error = 'Please select from dropdown'
    dv_retention.errorTitle = 'Invalid Retention Period'
    ws.add_data_validation(dv_retention)
    dv_retention.add(f'D10:D100')
    
    # Legal/Regulatory Basis (Column E)
    dv_legal = DataValidation(
        type="list",
        formula1='"Swiss FADP,EU GDPR,Swiss Code of Obligations (OR),Swiss Tax Law,EU ePrivacy Directive,Industry Standard (specify),Contractual Obligation,Legitimate Interest,Consent,Legal Obligation,Multiple Bases (specify),Other (specify in notes)"',
        allow_blank=True
    )
    dv_legal.error = 'Please select from dropdown'
    dv_legal.errorTitle = 'Invalid Legal Basis'
    ws.add_data_validation(dv_legal)
    dv_legal.add(f'E10:E100')
    
    # Status (Column F)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=True
    )
    dv_status.error = 'Please select from dropdown'
    dv_status.errorTitle = 'Invalid Status'
    ws.add_data_validation(dv_status)
    dv_status.add(f'F10:F100')
    
    # Risk Level (Column M)
    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    dv_risk.error = 'Please select from dropdown'
    dv_risk.errorTitle = 'Invalid Risk Level'
    ws.add_data_validation(dv_risk)
    dv_risk.add(f'M10:M100')
    
    # Budget Required (Column Q)
    dv_budget = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    dv_budget.error = 'Please select from dropdown'
    dv_budget.errorTitle = 'Invalid Budget Option'
    ws.add_data_validation(dv_budget)
    dv_budget.add(f'Q10:Q100')


def create_sheet_specific_validations(ws, sheet_type):
    """Create sheet-specific data validations."""
    
    if sheet_type == "sheet2":
        # Primary Storage Location (Column R)
        dv = DataValidation(
            type="list",
            formula1='"On-Premise,Cloud (IaaS),Cloud (PaaS),Cloud (SaaS),Hybrid,Third-Party"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Contains PII/SPI (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes - PII,Yes - SPI,Yes - Both,No"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
    
    elif sheet_type == "sheet3":
        # Retention Calculation Method (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Fixed Period,Event-Based,Hybrid"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Backup Retention Aligned (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes,No,Partial,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
    
    elif sheet_type == "sheet4":
        # Trigger Type (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Automatic,Manual,Semi-Automatic,Event-Based"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Trigger Frequency (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Real-time,Daily,Weekly,Monthly,Quarterly,Annual,Ad-hoc"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Legal Hold Check Integrated (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Automated,Yes - Manual,No,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
    
    elif sheet_type == "sheet5":
        # Legal Hold Notification Process (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Automated,Manual,Hybrid,None"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Hold Review Frequency (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Weekly,Monthly,Quarterly,Annual"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
    
    elif sheet_type == "sheet6":
        # GDPR/FADP Applicable (Column S)
        dv = DataValidation(
            type="list",
            formula1='"GDPR Only,FADP Only,Both,Neither"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')


# ==========================================================================
# SECTION 4: HELPER FUNCTIONS
# ==========================================================================

def apply_cell_style(cell, styles, style_type):
    """Apply style dictionary to a cell."""
    style = styles[style_type]
    cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


def create_header_row(ws, row, columns, styles):
    """Create header row with column definitions."""
    for col_letter, header_text, width in columns:
        cell = ws[f'{col_letter}{row}']
        cell.value = header_text
        apply_cell_style(cell, styles, 'header')
        ws.column_dimensions[col_letter].width = width


def create_data_rows(ws, start_row, num_rows, num_cols, styles):
    """Create yellow-highlighted data entry rows."""
    for row in range(start_row, start_row + num_rows):
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles, 'input_cell')


def create_checklist_section(ws, start_row, checklist_items, styles):
    """Create compliance checklist section."""
    # Section header
    ws.merge_cells(f'A{start_row}:T{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = "COMPLIANCE CHECKLIST - Mark ✅ or {XMARK}"
    apply_cell_style(cell, styles, 'subheader')
    
    # Checklist items
    current_row = start_row + 1
    for idx, item in enumerate(checklist_items, 1):
        # Checkbox column
        ws[f'A{current_row}'].value = f"{idx}."
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
        ws.column_dimensions['A'].width = 5
        
        # Item description
        ws.merge_cells(f'B{current_row}:R{current_row}')
        ws[f'B{current_row}'].value = item
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Status column
        ws[f'S{current_row}'].value = ""
        ws[f'S{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.column_dimensions['S'].width = 8
        
        # Notes column
        ws[f'T{current_row}'].value = ""
        ws[f'T{current_row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        current_row += 1
    
    return current_row


def create_reference_table(ws, start_row, table_title, headers, data, styles):
    """Create reference table with title and data."""
    # Table title
    num_cols = len(headers)
    end_col = get_column_letter(num_cols)
    ws.merge_cells(f'A{start_row}:{end_col}{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = table_title
    apply_cell_style(cell, styles, 'subheader')
    
    # Headers
    header_row = start_row + 1
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{header_row}']
        cell.value = header
        apply_cell_style(cell, styles, 'header')
    
    # Data rows
    current_row = header_row + 1
    for row_data in data:
        for col_idx, value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = value
            apply_cell_style(cell, styles, 'reference_cell')
        current_row += 1
    
    return current_row + 1  # Return next available row


# ==========================================================================
# SECTION 5: ASSESSMENT SHEET CREATOR (CORE FUNCTION)
# ==========================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, 
                            assessment_question, base_cols, extended_cols,
                            checklist_items, reference_tables, sheet_type):
    """
    Create standardized assessment sheet with:
    - Title and policy reference
    - Assessment question
    - Column headers
    - Data entry rows (13 rows, yellow-highlighted)
    - Compliance checklist
    - Reference tables
    - Exception/deviation block
    """
    
    # Row 1: Title
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = section_title
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    # Row 2: Policy Reference
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = f"Policy Reference: {policy_ref}"
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='left')
    
    # Row 3: Blank
    ws.row_dimensions[3].height = 5
    
    # Row 4-6: Assessment Question
    ws.merge_cells('A4:T6')
    cell = ws['A4']
    cell.value = f"ASSESSMENT QUESTION:\n{assessment_question}"
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[4].height = 50
    
    # Row 7: Instructions
    ws.merge_cells('A7:T7')
    cell = ws['A7']
    cell.value = "Complete the yellow-highlighted cells below. Use dropdowns where provided. Link evidence in Column N to Evidence Register sheet."
    cell.font = Font(name='Calibri', size=9, italic=True, color='0000FF')
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    # Row 8: Blank
    ws.row_dimensions[8].height = 5
    
    # Row 9: Column Headers (Base + Extended)
    all_columns = base_cols + extended_cols
    create_header_row(ws, 9, all_columns, styles)
    
    # Rows 10-22: Data Entry Rows (13 rows)
    create_data_rows(ws, 10, 13, len(all_columns), styles)
    
    # Row 23: Blank
    ws.row_dimensions[23].height = 10
    
    # Row 24: Blank
    ws.row_dimensions[24].height = 5
    
    # Rows 25+: Compliance Checklist
    next_row = create_checklist_section(ws, 25, checklist_items, styles)
    
    # Blank row after checklist
    ws.row_dimensions[next_row].height = 10
    next_row += 1
    
    # Reference Tables
    for table_title, headers, data in reference_tables:
        next_row = create_reference_table(ws, next_row, table_title, headers, data, styles)
    
    # Exception/Deviation Block
    ws.merge_cells(f'A{next_row}:T{next_row}')
    cell = ws[f'A{next_row}']
    cell.value = "EXCEPTIONS / DEVIATIONS"
    apply_cell_style(cell, styles, 'subheader')
    
    next_row += 1
    ws.merge_cells(f'A{next_row}:T{next_row+2}')
    cell = ws[f'A{next_row}']
    cell.value = "Document any exceptions or deviations from requirements here (e.g., data categories excluded from assessment, pending reviews, disputed retention periods):"
    cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[next_row].height = 60
    
    # Freeze panes at row 10 (headers stay visible)
    ws.freeze_panes = 'A10'
    
    # Apply validations
    create_base_validations(ws)
    create_sheet_specific_validations(ws, sheet_type)


# ==========================================================================
# SECTION 6: INSTRUCTIONS SHEET
# ==========================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # Title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ISMS-IMP-A.8.10.1 - Retention & Deletion Triggers Assessment"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    # Document Information Section
    current_row = 3
    ws[f'A{current_row}'].value = "Workbook Version:"
    ws[f'B{current_row}'].value = "1.0"
    ws[f'D{current_row}'].value = "Assessment Date:"
    ws[f'E{current_row}'].value = datetime.now().strftime('%d.%m.%Y')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Assessor Name:"
    ws[f'B{current_row}'].value = "[Enter Name]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    ws[f'D{current_row}'].value = "Organization:"
    ws[f'E{current_row}'].value = "[Enter Organization]"
    ws[f'E{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Review Period:"
    ws[f'B{current_row}'].value = "[e.g., Q4 2025]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 2
    
    # How to Use This Workbook
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "HOW TO USE THIS WORKBOOK"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    instructions = [
        "1. Complete assessment sheets 2-6 in order, filling in all yellow-highlighted cells",
        "2. Use dropdown menus where provided for consistency",
        "3. Link supporting evidence to the Evidence Register (Sheet 8) using Column N",
        "4. Mark checklist items as ✅ (compliant) or ❌ (non-compliant)",
        "5. Review the Summary Dashboard (Sheet 7) for overall compliance status",
        "6. Document any gaps with remediation plans and target completion dates",
        "7. Obtain three-level approval using the Approval Sign-Off sheet (Sheet 9)",
        "8. Review and update this assessment annually or when significant changes occur"
    ]
    
    for instruction in instructions:
        ws[f'A{current_row}'].value = instruction
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        current_row += 1
    
    current_row += 1
    
    # Color Coding Legend
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "COLOR CODING LEGEND"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    # Legend table
    legend_data = [
        ("Blue Header", "Column headers - Do not edit"),
        ("Yellow", "Data entry cells - Complete these fields"),
        ("Green", "Compliant status - Automated or manual entry"),
        ("Orange", "Partial compliance - Needs attention"),
        ("Red", "Non-compliant - Critical gap"),
        ("Gray", "Reference information - Read-only"),
        ("White", "Optional fields - Complete if relevant")
    ]
    
    for color_desc, usage in legend_data:
        ws[f'A{current_row}'].value = color_desc
        ws[f'B{current_row}'].value = usage
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        current_row += 1
    
    current_row += 1
    
    # Key Definitions
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "KEY DEFINITIONS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    definitions = [
        ("Data Category", "A classification of data based on content, purpose, or source (e.g., Customer Records, HR Files)"),
        ("Retention Period", "The duration for which data must or should be retained before deletion"),
        ("Deletion Trigger", "An event or condition that initiates data deletion (automated, manual, or event-based)"),
        ("Legal Hold", "Suspension of normal deletion due to litigation, investigation, or regulatory requirement"),
        ("Data Subject Request", "Request from individual to delete their personal data (GDPR Article 17 / FADP)"),
        ("Event-Based Retention", "Retention period calculated from a specific event (e.g., contract termination + 10 years)"),
        ("PII", "Personally Identifiable Information - data that can identify an individual"),
        ("SPI", "Special category Personal Information - sensitive data (health, biometric, etc.)")
    ]
    
    for term, definition in definitions:
        ws[f'A{current_row}'].value = term
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].value = definition
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 30
        current_row += 1
    
    current_row += 1
    
    # Regulatory References
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "REGULATORY REFERENCES"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    references = [
        f"{BULLET} GDPR Article 17: Right to Erasure ('Right to be Forgotten')",
        f"{BULLET} GDPR Article 5(1)(e): Storage Limitation Principle",
        f"{BULLET} Swiss FADP Article 6: Data Processing Principles",
        f"{BULLET} Swiss Code of Obligations (OR) Article 127, 958f: Record retention",
        f"{BULLET} ISO/IEC 27002:2022 Control A.8.10: Information Deletion",
        f"{BULLET} NIST SP 800-88 Rev. 1: Guidelines for Media Sanitization (reference only)"
    ]
    
    for ref in references:
        ws[f'A{current_row}'].value = ref
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        current_row += 1
    
    current_row += 1
    
    # Important Notes
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "IMPORTANT NOTES"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    notes = [
        "✓ This assessment is vendor-neutral. Document YOUR organization's specific implementations.",
        "✓ Link to related assessments: A.8.10.2 (Deletion Methods), A.8.10.3 (Third-Party), A.8.10.4 (Verification)",
        "✓ Annual review is required per ISMS-POL-A.8.10-S4 (Policy Governance)",
        "✓ Escalate critical gaps immediately to Data Protection Officer and Executive Management",
        "✓ Retain completed workbook per organizational records retention schedule (typically 3-7 years)"
    ]
    
    for note in notes:
        ws[f'A{current_row}'].value = note
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 25
        current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15


# ==========================================================================
# SECTION 7: EVIDENCE REGISTER
# ==========================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register with 100 rows."""
    
    # Title
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "Evidence Register - Supporting Documentation"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "Use this register to document all evidence supporting your retention and deletion trigger assessments. Reference evidence by ID (Column A) in the 'Evidence Reference' column (Column N) of assessment sheets."
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Row 3: Blank
    ws.row_dimensions[3].height = 5
    
    # Column Headers
    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Assessment Sheet", 20),
        ("C", "Related Data Category/Item", 30),
        ("D", "Evidence Type", 20),
        ("E", "Evidence Title/Description", 35),
        ("F", "File Location/Link", 40),
        ("G", "Date Created/Collected", 12),
        ("H", "Retention Period", 15),
        ("I", "Next Review Date", 12),
        ("J", "Owner/Custodian", 20),
        ("K", "Notes", 30)
    ]
    
    create_header_row(ws, 4, headers, styles)
    
    # Data rows (100 rows)
    for row in range(5, 105):
        for col_idx, (col_letter, _, _) in enumerate(headers, 1):
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles, 'input_cell')
    
    # Dropdowns for Evidence Type (Column D)
    dv_type = DataValidation(
        type="list",
        formula1='"Policy Document,Procedure Document,Screenshot/Print Screen,System Log Export,Configuration File,Email Communication,Meeting Minutes,Audit Report,Certificate (Deletion/Compliance),Contract/Agreement,Training Record,Test Result,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add('D5:D104')
    
    # Dropdowns for Assessment Sheet (Column B)
    dv_sheet = DataValidation(
        type="list",
        formula1='"Sheet 2: Data Category Registry,Sheet 3: Retention Schedule Compliance,Sheet 4: Deletion Trigger Configuration,Sheet 5: Legal Hold Management,Sheet 6: Data Subject Requests"',
        allow_blank=True
    )
    ws.add_data_validation(dv_sheet)
    dv_sheet.add('B5:B104')
    
    # Dropdowns for Retention Period (Column H)
    dv_retention = DataValidation(
        type="list",
        formula1='"1 year,3 years,5 years,7 years,10 years,Duration of related data retention + 1 year,Permanent"',
        allow_blank=True
    )
    ws.add_data_validation(dv_retention)
    dv_retention.add('H5:H104')
    
    # Freeze panes
    ws.freeze_panes = 'A5'


# ==========================================================================
# SECTION 8: APPROVAL SIGN-OFF
# ==========================================================================

def create_approval_signoff(ws, styles):
    """Create Approval Sign-Off workflow sheet."""
    
    # Title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "Three-Level Approval Workflow"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    # Document Control Section
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "DOCUMENT CONTROL"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    control_fields = [
        ("Assessment Period:", "[e.g., Q4 2025]"),
        ("Workbook Version:", "1.0"),
        ("Total Assessment Sheets Completed:", "5"),
        ("Overall Compliance % (from Dashboard):", "[Link to Summary Dashboard]"),
        ("Critical Gaps Identified:", "[Count from Summary]"),
        ("Assessment Completed By:", "[Name, Date]")
    ]
    
    for label, value in control_fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    current_row += 2
    
    # Level 1 Approval
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "LEVEL 1 APPROVAL - Technical/Operational"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = 'Role: Data Protection Officer / Privacy Officer / IT Security Manager'
    ws[f'A{current_row}'].font = Font(italic=True)
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 1}')
    ws[f'A{current_row}'].value = 'Approval Statement: "I confirm that this assessment accurately reflects our current retention schedules and deletion trigger implementations as of [Date]. All data categories have been reviewed, gaps have been identified, and remediation plans are in place."'
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 40
    current_row += 2
    
    approval_fields = [
        "Approver Name:",
        "Title/Role:",
        "Email:",
        "Review Date:",
        "Approval Status:",
        "Conditions/Comments:",
        "Signature:"
    ]
    
    for field in approval_fields:
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        if field == "Approval Status:":
            # Add dropdown
            dv = DataValidation(type="list", formula1=f'"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1
    
    current_row += 2
    
    # Level 2 Approval
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "LEVEL 2 APPROVAL - Management"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = 'Role: Chief Information Officer / Head of Compliance / Legal Counsel'
    ws[f'A{current_row}'].font = Font(italic=True)
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 1}')
    ws[f'A{current_row}'].value = 'Approval Statement: "I acknowledge the findings of this A.8.10.1 assessment and approve the proposed remediation plans. Resources will be allocated to address critical and high-risk gaps within the specified timelines."'
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 40
    current_row += 2
    
    for field in approval_fields:
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        if field == "Approval Status:":
            dv = DataValidation(type="list", formula1=f'"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1
    
    current_row += 2
    
    # Level 3 Approval
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "LEVEL 3 APPROVAL - Executive"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = 'Role: Chief Executive Officer / Chief Risk Officer / Board Delegate'
    ws[f'A{current_row}'].font = Font(italic=True)
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 1}')
    ws[f'A{current_row}'].value = 'Approval Statement: "This assessment has been reviewed at the executive level. The organization\'s retention and deletion trigger posture is [Acceptable/Needs Improvement/Unacceptable]. The Board/Executive Team has been briefed on critical gaps and remediation commitments."'
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 40
    current_row += 2
    
    for field in approval_fields[:-1]:  # All except Signature
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        if field == "Approval Status:":
            dv = DataValidation(type="list", formula1=f'"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1
    
    ws[f'A{current_row}'].value = "Executive Summary:"
    ws.merge_cells(f'B{current_row}:F{current_row + 2}')
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    ws[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 60
    current_row += 3
    
    ws[f'A{current_row}'].value = "Signature:"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 2
    
    # Next Steps
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "NEXT STEPS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    # Next Steps table
    ws[f'A{current_row}'].value = "Action Item"
    ws[f'B{current_row}'].value = "Responsible Party"
    ws[f'C{current_row}'].value = "Due Date"
    ws[f'D{current_row}'].value = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_cell_style(ws[f'{col}{current_row}'], styles, 'header')
    current_row += 1
    
    next_steps = [
        ("Implement remediation plans for critical gaps", "", "", "Pending"),
        ("Quarterly progress review on remediation", "", "", "Pending"),
        ("Annual re-assessment of A.8.10.1", "", f"{datetime.now().year + 1}-01-01", "Scheduled"),
        ("Update ISMS-POL-A.8.10 if needed", "", "", "Pending/Not Required"),
        ("Communicate changes to stakeholders", "", "", "Pending")
    ]
    
    for action, responsible, due, status in next_steps:
        ws[f'A{current_row}'].value = action
        ws[f'B{current_row}'].value = responsible
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'C{current_row}'].value = due
        ws[f'C{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'D{current_row}'].value = status
        current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


# ==========================================================================
# SECTION 9: SUMMARY DASHBOARD
# ==========================================================================

def create_summary_dashboard(ws, styles):
    """Create comprehensive summary dashboard."""
    
    # Title
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "A.8.10.1 Retention & Deletion Triggers - Compliance Dashboard"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    # Section 1: Overall Compliance Summary
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "OVERALL COMPLIANCE SUMMARY"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    # Headers
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{current_row}']
        cell.value = header
        apply_cell_style(cell, styles, 'header')
    current_row += 1
    
    # Compliance rows (formulas would go here in actual implementation)
    assessment_areas = [
        "Data Category Registry",
        "Retention Schedule Compliance",
        "Deletion Trigger Configuration",
        "Legal Hold Management",
        "Data Subject Requests"
    ]
    
    for area in assessment_areas:
        ws[f'A{current_row}'].value = area
        ws[f'B{current_row}'].value = "[Formula: COUNT]"
        ws[f'C{current_row}'].value = "[Formula: COUNTIF Compliant]"
        ws[f'D{current_row}'].value = "[Formula: COUNTIF Partial]"
        ws[f'E{current_row}'].value = "[Formula: COUNTIF Non-Compliant]"
        ws[f'F{current_row}'].value = "[Formula: COUNTIF N/A]"
        ws[f'G{current_row}'].value = "[Formula: %]"
        current_row += 1
    
    # Overall row
    ws[f'A{current_row}'].value = "OVERALL A.8.10.1"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'].value = "[SUM]"
    ws[f'C{current_row}'].value = "[SUM]"
    ws[f'D{current_row}'].value = "[SUM]"
    ws[f'E{current_row}'].value = "[SUM]"
    ws[f'F{current_row}'].value = "[SUM]"
    ws[f'G{current_row}'].value = "[Overall %]"
    current_row += 2
    
    # Compliance Thresholds
    ws[f'A{current_row}'].value = "Compliance Thresholds:"
    ws[f'B{current_row}'].value = "≥90% = 🟢 Excellent | 70-89% = 🟡 Needs Improvement | <70% = 🔴 Critical"
    ws[f'B{current_row}'].font = Font(italic=True, size=9)
    current_row += 2
    
    # Section 2: Critical Gaps
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    gap_headers = ["Data Category/Item", "Gap Description", "Risk Level", "Target Completion", "Owner"]
    for col_idx, header in enumerate(gap_headers, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{current_row}']
        cell.value = header
        apply_cell_style(cell, styles, 'header')
    current_row += 1
    
    ws[f'A{current_row}'].value = "[Auto-populated from assessment sheets where Status = Non-Compliant AND Risk = Critical/High]"
    ws[f'A{current_row}'].font = Font(italic=True, size=9)
    current_row += 3
    
    # Section 3: Retention Schedule Coverage
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "RETENTION SCHEDULE COVERAGE"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    metrics = [
        ("Total Data Categories Identified", "[Count from Sheet 2]", "N/A", "ℹ️ Info"),
        ("Data Categories with Defined Retention", "[Count from Sheet 3]", "100%", "[Status]"),
        ("Data Categories with Legal Basis", "[Count from Sheet 3]", "100%", "[Status]"),
        ("Event-Based Retention Categories", "[Count]", "N/A", "ℹ️ Info"),
        ("Categories with PII/SPI", "[Count]", "N/A", "ℹ️ Info"),
        ("Categories Pending Review", "[Count overdue]", "0", "[Status]")
    ]
    
    ws[f'A{current_row}'].value = "Metric"
    ws[f'B{current_row}'].value = "Count/Percentage"
    ws[f'C{current_row}'].value = "Target"
    ws[f'D{current_row}'].value = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_cell_style(ws[f'{col}{current_row}'], styles, 'header')
    current_row += 1
    
    for metric, count, target, status in metrics:
        ws[f'A{current_row}'].value = metric
        ws[f'B{current_row}'].value = count
        ws[f'C{current_row}'].value = target
        ws[f'D{current_row}'].value = status
        current_row += 1
    
    current_row += 1
    
    # Section 4: Deletion Trigger Effectiveness
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "DELETION TRIGGER EFFECTIVENESS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    trigger_metrics = [
        ("Categories with Automated Triggers", "[Count from Sheet 4]", "≥70%", "[Status]"),
        ("Categories with Manual Triggers Only", "[Count]", "≤30%", "[Status]"),
        ("Triggers with Legal Hold Integration", "[Count]", "100%", "[Status]"),
        ("Trigger Test Failures (Last 12 Months)", "[Manual entry]", "0", "[Status]"),
        ("Average Trigger Execution Success Rate", "[Calculated]", "≥99%", "[Status]")
    ]
    
    ws[f'A{current_row}'].value = "Metric"
    ws[f'B{current_row}'].value = "Count/Percentage"
    ws[f'C{current_row}'].value = "Target"
    ws[f'D{current_row}'].value = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_cell_style(ws[f'{col}{current_row}'], styles, 'header')
    current_row += 1
    
    for metric, count, target, status in trigger_metrics:
        ws[f'A{current_row}'].value = metric
        ws[f'B{current_row}'].value = count
        ws[f'C{current_row}'].value = target
        ws[f'D{current_row}'].value = status
        current_row += 1
    
    current_row += 1
    
    # Section 5: Legal Hold Management
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "LEGAL HOLD MANAGEMENT METRICS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    hold_metrics = [
        ("Active Legal Holds", "[SUM from Sheet 5]", "Current snapshot"),
        ("Average Hold Duration (Days)", "[Calculated]", "Monitor for stale holds"),
        ("Legal Hold Breaches (Last 12 Months)", "[Manual entry]", "Target: 0"),
        ("Systems Without Legal Hold Capability", "[Count Non-Compliant]", "Risk assessment required"),
        ("Last Legal Hold Process Test", "[Date entry]", "Annual testing required")
    ]
    
    ws[f'A{current_row}'].value = "Metric"
    ws[f'B{current_row}'].value = "Value"
    ws[f'C{current_row}'].value = "Notes"
    for col in ['A', 'B', 'C']:
        apply_cell_style(ws[f'{col}{current_row}'], styles, 'header')
    current_row += 1
    
    for metric, value, notes in hold_metrics:
        ws[f'A{current_row}'].value = metric
        ws[f'B{current_row}'].value = value
        ws[f'C{current_row}'].value = notes
        current_row += 1
    
    current_row += 1
    
    # Section 6: Data Subject Request Performance
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "DATA SUBJECT REQUEST PERFORMANCE (GDPR/FADP)"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    dsr_metrics = [
        ("Total Requests (Last 12 Months)", "[SUM from Sheet 6]", "N/A", "ℹ️ Info"),
        ("Average Response Time (Days)", "[AVERAGE]", "≤30 days", "[Status]"),
        ("Requests Exceeding 30 Days", "[COUNTIF >30]", "0", "[Status]"),
        ("Systems Unable to Delete Data Subject Data", "[Count]", "0", "[Status]"),
        ("Exception Rate (% not fully deleted)", "[Calculated]", "<20%", "[Status]"),
        ("GDPR/FADP Applicable Systems", "[Count]", "N/A", "ℹ️ Info")
    ]
    
    ws[f'A{current_row}'].value = "Metric"
    ws[f'B{current_row}'].value = "Value"
    ws[f'C{current_row}'].value = "Target"
    ws[f'D{current_row}'].value = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_cell_style(ws[f'{col}{current_row}'], styles, 'header')
    current_row += 1
    
    for metric, value, target, status in dsr_metrics:
        ws[f'A{current_row}'].value = metric
        ws[f'B{current_row}'].value = value
        ws[f'C{current_row}'].value = target
        ws[f'D{current_row}'].value = status
        current_row += 1
    
    current_row += 1
    
    # Section 7: Risk Summary
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "RISK SUMMARY"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Risk Level"
    ws[f'B{current_row}'].value = "Count of Open Gaps"
    ws[f'C{current_row}'].value = "% of Total Gaps"
    ws[f'D{current_row}'].value = "Priority Action"
    for col in ['A', 'B', 'C', 'D']:
        apply_cell_style(ws[f'{col}{current_row}'], styles, 'header')
    current_row += 1
    
    risk_levels = [
        ("Critical", "[COUNTIF Critical]", "[%]", "Immediate escalation"),
        ("High", "[COUNTIF High]", "[%]", "Executive attention"),
        ("Medium", "[COUNTIF Medium]", "[%]", "Planned remediation"),
        ("Low", "[COUNTIF Low]", "[%]", "Monitor")
    ]
    
    for level, count, pct, action in risk_levels:
        ws[f'A{current_row}'].value = level
        ws[f'B{current_row}'].value = count
        ws[f'C{current_row}'].value = pct
        ws[f'D{current_row}'].value = action
        current_row += 1
    
    current_row += 1
    
    # Section 8: Executive Summary
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "EXECUTIVE SUMMARY & RECOMMENDATIONS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Overall A.8.10.1 Maturity Level:"
    ws[f'B{current_row}'].value = "[Emerging / Developing / Established / Optimized]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 2
    
    ws[f'A{current_row}'].value = "Key Strengths:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    for i in range(1, 4):
        ws[f'A{current_row}'].value = f"{i}."
        ws[f'B{current_row}'].value = "[Example strength]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    current_row += 1
    ws[f'A{current_row}'].value = "Critical Improvement Areas:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    for i in range(1, 4):
        ws[f'A{current_row}'].value = f"{i}."
        ws[f'B{current_row}'].value = "[Critical gap]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    current_row += 1
    ws[f'A{current_row}'].value = "Top 3 Recommendations:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    for i in range(1, 4):
        ws[f'A{current_row}'].value = f"{i}."
        ws[f'B{current_row}'].value = "[Priority action with timeline]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1
    
    current_row += 1
    ws[f'A{current_row}'].value = "Next Review Date:"
    ws[f'B{current_row}'].value = "[Typically annual, e.g., Q4 2026]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15


# ==========================================================================
# SECTION 10: DOMAIN-SPECIFIC SHEET CREATORS
# ==========================================================================

def create_sheet2_data_category_registry(ws, styles):
    """Create Sheet 2: Data Category Registry."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet2()
    
    checklist = [
        "All business units have been consulted for data categories",
        "Data inventory includes operational, HR, financial, and customer data",
        "Each data category has a designated Business Owner",
        "Storage locations (on-premise, cloud, third-party) are documented",
        "PII and SPI data categories are specifically identified",
        "Volume/scale of data is estimated for each category",
        "Data classification levels are assigned per organizational policy",
        "Temporary/ephemeral data categories are included",
        "Legacy systems and archived data are included in inventory",
        "Cross-border data transfers are noted (if applicable)",
        "Data category registry is reviewed at least annually",
        "New business processes trigger data category reassessment",
        "Shadow IT data categories have been investigated",
        "M&A due diligence includes data category mapping",
        "Data categories align with Records Retention Schedule"
    ]
    
    reference_tables = [
        (
            "Common Data Categories",
            ["Category Example", "Typical Retention", "Common Legal Basis"],
            [
                ["Employee Personnel Files", "10 years post-employment", "Swiss OR, Tax Law"],
                ["Customer Contracts", "10 years post-termination", "Swiss OR (Art. 127)"],
                ["Financial Records", "10 years", "Swiss Tax Law"],
                ["Marketing Consent Records", "Until consent withdrawn + 1 year", "GDPR, FADP"],
                ["Access Logs", "1 year", "Security requirement"],
                ["Email Communications", "1-7 years (varies)", "Business need"],
                ["Customer Support Tickets", "3 years post-resolution", "Contractual"],
                ["Product Telemetry", "90 days - 2 years", "Legitimate interest"],
                ["CCTV Footage", "30-90 days", "Security/legal requirement"],
                ["Website Analytics", "26 months (GDPR default)", "Consent/Legitimate interest"],
                ["Backup Data", "Aligned with active retention", "Same as source data"],
                ["Development/Test Data", "Until project end + 90 days", "Business need"],
                ["Audit Trails", "7-10 years", "Regulatory requirement"],
                ["HR Recruitment Records", "6 months post-decision", "GDPR, FADP"],
                ["Health & Safety Records", "10 years", "Legal obligation"]
            ]
        ),
        (
            "Storage Location Implications",
            ["Location Type", "Deletion Complexity", "Key Considerations"],
            [
                ["On-Premise", "Medium", "Full control, hardware lifecycle"],
                ["Cloud (IaaS)", "Medium-High", "Snapshot management, API deletion"],
                ["Cloud (PaaS)", "High", "Platform-specific deletion methods"],
                ["Cloud (SaaS)", "Very High", "Vendor-dependent, see A.8.10.3"],
                ["Hybrid", "Very High", "Coordination across environments"],
                ["Third-Party", "Very High", "Contractual deletion, see A.8.10.3"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 2: Data Category Registry",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.1",
        "Have we identified and documented ALL data categories we process, including their characteristics, ownership, and storage locations?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet2"
    )


def create_sheet3_retention_schedule(ws, styles):
    """Create Sheet 3: Retention Schedule Compliance."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet3()
    
    checklist = [
        "Every data category has a defined retention period",
        "Retention periods have documented legal/regulatory basis",
        "Swiss FADP compliance verified for all PII categories",
        "GDPR compliance verified where applicable",
        "Retention periods do not exceed legal maximums",
        "Retention periods meet legal minimums (e.g., tax records)",
        "Event-based retention triggers are clearly defined",
        "Business justification documented for non-regulatory retention",
        "Conflicting retention requirements are resolved and documented",
        "Retention schedule approved by Legal/Compliance team",
        "Backup retention aligns with active data retention",
        "'Permanent' retention is justified and approved",
        "Retention periods reviewed annually",
        "Changes to retention periods follow change management process",
        "Data subjects are informed of retention periods (GDPR/FADP)"
    ]
    
    reference_tables = [
        (
            "Swiss Legal Retention Requirements",
            ["Data Type", "Minimum Retention", "Legal Basis", "Maximum Deletion"],
            [
                ["Accounting Records", "10 years", "Swiss OR Art. 958f", "After 10 years"],
                ["Tax Records", "10 years", "Swiss Tax Law", "After 10 years"],
                ["Employee Records", "10 years", "Swiss OR Art. 127", "After 10 years (FADP)"],
                ["Contracts", "10 years", "Swiss OR Art. 127", "After 10 years"],
                ["Payroll Records", "10 years", "Swiss Tax/Social Law", "After 10 years"],
                ["Corporate Minutes", "Permanent", "Swiss OR Art. 695", "N/A"],
                ["Share Register", "Permanent", "Swiss OR Art. 686", "N/A"],
                ["Correspondence", "No minimum", "Business need", "FADP: when no longer needed"],
                ["Marketing Data", "No minimum", "Consent-based", "FADP: consent withdrawal"],
                ["CCTV Footage", "No minimum", "Proportionality", "FADP: typically 30-90 days"]
            ]
        ),
        (
            "Retention Calculation Methods",
            ["Method", "Definition", "Example"],
            [
                ["Fixed Period", "Set duration from creation/collection", "7 years from date of invoice"],
                ["Event-Based", "Duration starts at specific event", "10 years from contract termination"],
                ["Hybrid", "Combination of fixed + event", "3 years OR until litigation resolved"],
                ["Indefinite (Permanent)", "No planned deletion", "Corporate charter documents"],
                ["Until Consent Withdrawn", "Active until user action", "Marketing email list"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 3: Retention Schedule Compliance",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.2",
        "Does each data category have a documented retention period with a valid legal or business justification?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet3"
    )


def create_sheet4_deletion_triggers(ws, styles):
    """Create Sheet 4: Deletion Trigger Configuration."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet4()
    
    checklist = [
        "Automated deletion triggers implemented for high-volume data",
        "Manual deletion procedures documented for low-volume data",
        "Trigger execution is logged and auditable",
        "Failed deletion attempts are alerted and investigated",
        "Legal hold checks are integrated into trigger logic",
        "Backup deletion triggers aligned with active data triggers",
        "Cloud/SaaS deletion triggers verified (see A.8.10.3)",
        "Event-based triggers are monitored for event occurrence",
        "Deletion triggers tested at least annually",
        "Trigger failures have defined escalation procedures",
        "Business Owner approval required for manual triggers (if policy requires)",
        "Deletion triggers respect data dependencies (referential integrity)",
        "Triggers account for data in transit (queues, caches)",
        "Deletion trigger schedule published and communicated",
        "Trigger configuration changes follow change management"
    ]
    
    reference_tables = [
        (
            "Trigger Types and Use Cases",
            ["Trigger Type", "Best For", "Advantages", "Disadvantages"],
            [
                ["Automatic", "High-volume, structured data", "Consistent, no human error", "Requires robust testing"],
                ["Manual", "Low-volume, sensitive data", "Controlled, deliberate", "Human error risk, slower"],
                ["Semi-Automatic", "Medium volume, needs approval", "Balance control and efficiency", "Workflow complexity"],
                ["Event-Based", "Contract-dependent data", "Precise, legally aligned", "Requires event tracking"]
            ]
        ),
        (
            "Common Deletion Trigger Patterns",
            ["Pattern", "Description", "Example Implementation"],
            [
                ["Time-Based Cron", "Scheduled deletion job", "Daily at 02:00 UTC, delete records > 7 years"],
                ["Database TTL", "Built-in expiration", "MongoDB TTL index, Redis EXPIRE"],
                ["Lifecycle Policy", "Cloud-native deletion", "AWS S3 Lifecycle, Azure Blob Lifecycle"],
                ["Workflow Approval", "Human-in-the-loop", "Jira ticket → approval → script execution"],
                ["Event Listener", "Reactive deletion", "Contract.status = terminated → 90-day timer → delete"],
                ["Batch Processing", "Periodic bulk deletion", "Monthly job: identify + delete expired records"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 4: Deletion Trigger Configuration",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.3",
        "Do we have effective and reliable deletion triggers that ensure data is deleted when retention periods expire?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet4"
    )


def create_sheet5_legal_hold(ws, styles):
    """Create Sheet 5: Legal Hold Management."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet5()
    
    checklist = [
        "Legal hold procedures are documented and approved",
        "Legal/Compliance team has authority to initiate holds",
        "IT/Data teams receive timely notification of holds",
        "Hold notification includes specific data categories and date ranges",
        "Automated deletion triggers check for active holds",
        "Legal hold data is segregated or flagged in systems",
        "Legal hold registry is maintained (who, what, when, why)",
        "Business Owners are notified when their data is under hold",
        "Periodic review of active holds (at least quarterly)",
        "Hold release procedures are documented",
        "Post-hold deletion is executed per normal retention schedule",
        "Legal hold training provided to relevant staf",
        "Legal hold breaches (accidental deletion) are reported and investigated",
        "Legal hold scope is clearly defined (not overly broad)",
        "Legal hold process tested at least annually"
    ]
    
    reference_tables = [
        (
            "Legal Hold Triggers",
            ["Trigger Event", "Typical Response Time", "Hold Scope"],
            [
                ["Litigation Notice", "Within 24 hours", "All relevant data for case"],
                ["Regulatory Investigation", "Within 48 hours", "Specified data types/periods"],
                ["Internal Investigation", "Within 72 hours", "Specific employees/systems"],
                ["M&A Due Diligence", "Within 1 week", "All corporate records"],
                ["Employment Dispute", "Within 48 hours", "Employee records + communications"],
                ["IP Dispute", "Within 72 hours", "Development records, communications"],
                ["Data Breach Investigation", "Immediate", "Affected systems/logs"],
                ["Audit Notification", "Within 1 week", "Records relevant to audit scope"]
            ]
        ),
        (
            "Hold Management Best Practices",
            ["Practice", "Purpose", "Implementation"],
            [
                ["Centralized Hold Registry", "Single source of truth", "Legal hold database/system"],
                ["Automated Flagging", "Prevent accidental deletion", "System flags or separate storage"],
                ["Regular Hold Reviews", "Release unnecessary holds", "Quarterly review meetings"],
                ["Clear Release Process", "Controlled return to normal deletion", "Legal approval required"],
                ["Documentation Requirements", "Audit trail", "Who, what, when, why, how long"],
                ["Scope Definition", "Avoid over-preservation", "Specific data types, date ranges"],
                ["Cost Tracking", "Understand hold impact", "Storage costs, resource time"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 5: Legal Hold Management",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.4",
        "Do we have robust processes to identify, communicate, and manage legal holds to prevent premature deletion of data under hold?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet5"
    )


def create_sheet6_data_subject_requests(ws, styles):
    """Create Sheet 6: Data Subject Requests (GDPR/FADP)."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet6()
    
    checklist = [
        "Data subject request procedure is documented",
        "Request intake channel is clearly communicated (e.g., privacy@company.ch)",
        "Identity verification process prevents fraudulent requests",
        "Request acknowledgment within 5 business days (GDPR/FADP)",
        "Requests fulfilled within 30 days (extendable to 60 days with justification)",
        "Search capability across all systems holding personal data",
        "Ability to distinguish between deletion obligations and exceptions",
        "Third-party/cloud systems included in deletion scope (see A.8.10.3)",
        "Backup deletion addressed in response (or timeline communicated)",
        "Request tracking system in place (ticket system, CRM, etc.)",
        "Responses documented (what was deleted, what was retained, why)",
        "Data subjects informed of retention exceptions (e.g., legal obligations)",
        "Escalation procedure for complex requests",
        "Staff training on data subject request handling",
        "Request metrics tracked and reported to DPO/Privacy Officer"
    ]
    
    reference_tables = [
        (
            "GDPR Article 17 Deletion Obligations",
            ["Condition for Deletion", "Organization Must Delete If...", "Example"],
            [
                ["No longer necessary", "Data no longer needed for original purpose", "Marketing data after consent withdrawn"],
                ["Consent withdrawn", "Individual withdraws consent (no other legal basis)", "Newsletter subscription cancellation"],
                ["Objection to processing", "Individual objects + no overriding legitimate grounds", "Profiling/direct marketing objection"],
                ["Unlawful processing", "Data was processed illegally", "Processing without valid legal basis"],
                ["Legal obligation", "Law requires deletion", "Court order, regulatory requirement"],
                ["Child's data", "Data collected from child w/o proper consent", "Social media account created under age 13"]
            ]
        ),
        (
            "GDPR Article 17 Deletion Exceptions",
            ["Exception", "Organization Can Retain If...", "Example"],
            [
                ["Freedom of expression", "Exercise of freedom of expression/information", "Journalistic/academic/artistic purposes"],
                ["Legal obligation", "Processing necessary for legal compliance", "Tax records (10-year retention)"],
                ["Public interest", "Public health, scientific research (safeguarded)", "Anonymized medical research data"],
                ["Legal claims", "Establishment, exercise, defense of legal claims", "Litigation hold, contract disputes"],
                ["Archiving in public interest", "Historical, statistical, scientific research (safeguarded)", "National archives, public health studies"]
            ]
        ),
        (
            "Request Response Workflow",
            ["Step", "Action", "Timeframe", "Responsible Role"],
            [
                ["1. Receive", "Log request in tracking system", "Day 0", "Privacy Team"],
                ["2. Acknowledge", "Send acknowledgment to data subject", "Within 5 business days", "Privacy Team"],
                ["3. Verify Identity", "Validate requestor identity", "Within 7 days", "Privacy Team"],
                ["4. Assess Scope", "Identify all systems with data subject's data", "Within 10 days", "IT + Business Owners"],
                ["5. Check Exceptions", "Determine if deletion exceptions apply", "Within 15 days", "Legal/Compliance"],
                ["6. Execute Deletion", "Delete data across all systems", "Within 25 days", "IT + Third-Parties"],
                ["7. Document", "Record what was deleted and retained", "Day 28", "Privacy Team"],
                ["8. Respond", "Inform data subject of outcome", "Day 30", "Privacy Team"],
                ["9. Follow-Up", "Backup deletion (if applicable)", "Within 90 days", "IT"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 6: Data Subject Requests (GDPR/FADP)",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.5",
        "Can we effectively identify, validate, and execute deletion requests from data subjects within required timeframes?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet6"
    )


# ==========================================================================
# SECTION 11: REFERENCE DATA
# ==========================================================================

# (Checklists and reference tables are defined inline in sheet creators above)


# ==========================================================================
# SECTION 12: MAIN EXECUTION
# ==========================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    try:
        logger.info("=" * 78)
        logger.info("ISMS-IMP-A.8.10.1 - Retention & Deletion Triggers Assessment Generator")
        logger.info("ISO/IEC 27001:2022 Control A.8.10: Information Deletion")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/9] Creating Instructions & Legend...")
        create_instructions_sheet(wb["Instructions & Legend"], styles)

        logger.info("[2/9] Creating Sheet 2: Data Category Registry...")
        create_sheet2_data_category_registry(wb["2. Data Category Registry"], styles)

        logger.info("[3/9] Creating Sheet 3: Retention Schedule Compliance...")
        create_sheet3_retention_schedule(wb["3. Retention Schedule Compliance"], styles)

        logger.info("[4/9] Creating Sheet 4: Deletion Trigger Configuration...")
        create_sheet4_deletion_triggers(wb["4. Deletion Trigger Configuration"], styles)

        logger.info("[5/9] Creating Sheet 5: Legal Hold Management...")
        create_sheet5_legal_hold(wb["5. Legal Hold Management"], styles)

        logger.info("[6/9] Creating Sheet 6: Data Subject Requests...")
        create_sheet6_data_subject_requests(wb["6. Data Subject Requests"], styles)

        logger.info("[7/9] Creating Summary Dashboard...")
        create_summary_dashboard(wb["Summary Dashboard"], styles)

        logger.info("[8/9] Creating Evidence Register (100 rows)...")
        create_evidence_register(wb["Evidence Register"], styles)

        logger.info("[9/9] Creating Approval Sign-Off (3-level workflow)...")
        create_approval_signoff(wb["Approval Sign-Off"], styles)

        filename = f"ISMS-IMP-A.8.10.1_Retention_Deletion_Triggers_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("=" * 78)
        logger.info("SUCCESS: %s", filename)
        logger.info("Workbook Structure:")
        logger.info("  - Instructions & Legend - Complete usage guide")
        logger.info("  - Sheet 2: Data Category Registry - Complete data inventory")
        logger.info("  - Sheet 3: Retention Schedule Compliance - Legal/regulatory alignment")
        logger.info("  - Sheet 4: Deletion Trigger Configuration - Automated and manual triggers")
        logger.info("  - Sheet 5: Legal Hold Management - Hold procedures and active holds")
        logger.info("  - Sheet 6: Data Subject Requests - GDPR Article 17 / FADP compliance")
        logger.info("  - Summary Dashboard - Executive overview with KPIs")
        logger.info("  - Evidence Register - 100 rows for supporting documentation")
        logger.info("  - Approval Sign-Off - 3-level approval workflow")
        logger.info("Key Features: Vendor-neutral, 13 data entry rows, compliance checklists")
        logger.info("Next Steps: Open workbook, complete yellow fields, review dashboard")
        logger.info("Related: A.8.10.2 (Deletion Methods), A.8.10.3 (Third-Party)")
        logger.info("=" * 78)
        return 0
    except Exception as e:
        logger.error("Failed to generate workbook: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
