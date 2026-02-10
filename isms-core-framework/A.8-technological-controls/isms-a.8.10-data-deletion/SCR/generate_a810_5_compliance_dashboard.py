#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# Licensed under AGPL-3.0-or-later with commercial licensing option
#
# This file is part of the ISMS Compliance Framework
# See /LICENSE for full terms and /LICENSES/COMMERCIAL.md for commercial options
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.10.5 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.10: Information Deletion
Assessment Domain 5 of 5: Executive Compliance Dashboard and Consolidation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment workbook structures, cell reference
mappings, and dashboard requirements.

Key customization areas:
1. Cell reference mappings (CRITICAL - verify actual cell locations in workbooks)
2. Dashboard KPI calculations (adapt to your compliance thresholds)
3. Executive summary formatting (match your reporting standards)
4. Risk scoring methodologies (aligned with your risk framework)
5. Remediation tracking integration (based on your project management tools)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

**CRITICAL WARNING - CELL REFERENCE VERIFICATION:**
This dashboard uses EXTERNAL WORKBOOK FORMULA LINKS to pull data from the four
assessment workbooks (A.8.10.1 through A.8.10.4). The cell references in this
script are VERIFIED against the actual generator scripts, but you MUST verify
they match your generated workbooks:

Verified Cell References:
- A.8.10.1: 'Summary Dashboard'!G10 (Compliance %), E10 (Non-Compliant Count)
- A.8.10.2: 'Summary Dashboard'!G10 (Compliance %), E10 (Non-Compliant Count)
- A.8.10.3: 'Summary Dashboard'!G10 (Compliance %), E10 (Non-Compliant Count)
- A.8.10.4: 'Verification Dashboard'!B9 (Compliance %), B8 (Critical Gaps)

NOTE: A.8.10.4 uses DIFFERENT sheet name ('Verification Dashboard' not 'Summary Dashboard')

If you modify the assessment workbook structures, you MUST update these references.

Reference Pattern: Based on ISMS-A.8.10 Information Deletion Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel dashboard for consolidating deletion
compliance metrics across all A.8.10 assessment domains (Retention Triggers,
Deletion Methods, Third-Party Deletion, Verification & Evidence) providing
executive-level visibility into information deletion compliance posture.

**Purpose:**
Provides executive leadership, CISO, Data Protection Officer, and auditors with
consolidated view of organization's information deletion compliance status,
critical gaps, and remediation progress.

**Dashboard Scope:**
- Executive summary dashboard with overall compliance percentage
- Domain-specific compliance breakdowns (4 assessment domains)
- Critical gap identification and risk scoring
- Consolidated evidence master index across all domains
- Remediation roadmap and action item tracking
- CISO/DPO approval workflow
- KPI tracking and trend analysis
- Integration with all A.8.10 assessment workbooks via external formulas

**Generated Workbook Structure:**
1. Instructions & Legend - Dashboard usage guidance, interpretation, assumptions
2. Executive Summary - Overall compliance status, critical metrics, risk summary
3. Domain 1 Summary - Retention & Deletion Triggers compliance details
4. Domain 2 Summary - Deletion Methods compliance details
5. Domain 3 Summary - Third-Party & Cloud Deletion compliance details
6. Domain 4 Summary - Verification & Evidence compliance details
7. Consolidated Gap Analysis - All critical/high gaps across domains
8. Risk Register - Deletion-related risks and mitigations
9. Remediation Roadmap - Action items, timelines, ownership
10. Evidence Master Index - Comprehensive evidence inventory
11. KPI Dashboard - Trend analysis and performance metrics
12. CISO/DPO Approval - Executive sign-off and attestation

**Key Features:**
- External workbook formula links for automatic data updates
- Conditional formatting for compliance status visualization
- Automated gap consolidation from all assessment domains
- Risk scoring based on gap severity and data sensitivity
- Evidence traceability across all domains
- Executive approval workflow
- KPI trend tracking over time
- Integration with all four A.8.10 assessment workbooks

**Integration:**
This dashboard consolidates data from all four A.8.10 assessment workbooks:
- ISMS-IMP-A.8.10.1: Retention & Deletion Triggers
- ISMS-IMP-A.8.10.2: Deletion Methods
- ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion
- ISMS-IMP-A.8.10.4: Verification & Evidence

**Data Integration Method:**
Uses EXTERNAL WORKBOOK FORMULA REFERENCES for real-time data updates:
```
='[ISMS-IMP-A.8.10.1.xlsx]Summary Dashboard'!$G$10
```
When source workbooks are updated, dashboard reflects changes after "Update Links".

**Alternative Integration:**
For environments where external links don't work, use the companion script:
`consolidate_a810_dashboard_data.py` to directly copy data values.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

File Dependencies (Must be in same directory as dashboard):
    - ISMS-IMP-A.8.10.1.xlsx (Retention & Deletion Triggers - normalized)
    - ISMS-IMP-A.8.10.2.xlsx (Deletion Methods - normalized)
    - ISMS-IMP-A.8.10.3.xlsx (Third-Party & Cloud - normalized)
    - ISMS-IMP-A.8.10.4.xlsx (Verification & Evidence - normalized)

NOTE: Use normalized file names (run normalize_assessment_files_a810.py first)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a810_5_compliance_dashboard.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a810_5_compliance_dashboard.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a810_5_compliance_dashboard.py --date 20250124

Output:
    File: ISMS_A_8_10_5_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. **CRITICAL:** Place dashboard in SAME DIRECTORY as normalized assessment files
    2. Run: python3 normalize_assessment_files_a810.py (to create normalized names)
    3. Open dashboard in Excel
    4. Excel will prompt: "Update Links?" → Click "Update"
    5. Verify all external formulas display data (not #REF! errors)
    6. Review executive summary and compliance metrics
    7. Validate consolidated gap analysis
    8. Review risk register and remediation roadmap
    9. Obtain CISO/DPO approval and sign-off
    10. Distribute to stakeholders for remediation tracking

Troubleshooting External Links:
    - If #REF! errors appear: Check source files are in same directory
    - If values don't update: Click Data → Refresh All
    - If links break: Use consolidate_a810_dashboard_data.py to copy data directly

Alternative Workflow (No External Links):
    1. Generate dashboard: python3 generate_a810_5_compliance_dashboard.py
    2. Consolidate data: python3 consolidate_a810_dashboard_data.py dashboard.xlsx
    3. Open dashboard (data is copied, not linked)
    4. Re-run consolidation script when source data changes

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.10
Assessment Domain:    5 of 5 (Executive Compliance Dashboard and Consolidation)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.10: Information Deletion Policy (Governance)
    - ISMS-IMP-A.8.10.1: Retention & Deletion Triggers Assessment (Domain 1)
    - ISMS-IMP-A.8.10.2: Deletion Methods Assessment (Domain 2)
    - ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion Assessment (Domain 3)
    - ISMS-IMP-A.8.10.4: Verification & Evidence Assessment (Domain 4)
    - ISMS-IMP-A.8.10.5: Compliance Dashboard Guide

Related Scripts:
    - generate_a810_1_retention_triggers.py (Domain 1 generator)
    - generate_a810_2_deletion_methods.py (Domain 2 generator)
    - generate_a810_3_third_party_cloud.py (Domain 3 generator)
    - generate_a810_4_verification_evidence.py (Domain 4 generator)
    - normalize_assessment_files_a810.py (File normalization for external linking)
    - consolidate_a810_dashboard_data.py (Alternative data consolidation)
    - excel_sanity_check_a810.py (Quality assurance validation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full dashboard framework per ISMS-IMP-A.8.10.5 specification
    - Supports external workbook linking for automatic data updates
    - Verified cell references against actual assessment generator scripts
    - Integrated executive summary, gap analysis, risk register, remediation roadmap

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**CRITICAL: Cell Reference Verification**
This script contains VERIFIED cell references that were checked against the actual
assessment generator scripts (A.8.10.1 through A.8.10.4). However, if you modify
the assessment workbook structures, you MUST update the CELL_MAPPINGS dictionary
in this script to match your changes.

Current VERIFIED mappings:
```python
CELL_MAPPINGS = {
    'a810_1': {
        'sheet': 'Summary Dashboard',
        'compliance_pct': 'G10',
        'non_compliant': 'E10'
    },
    'a810_2': {
        'sheet': 'Summary Dashboard',
        'compliance_pct': 'G10',
        'non_compliant': 'E10'
    },
    'a810_3': {
        'sheet': 'Summary Dashboard',
        'compliance_pct': 'G10',
        'non_compliant': 'E10'
    },
    'a810_4': {
        'sheet': 'Verification Dashboard',  # DIFFERENT sheet name!
        'compliance_pct': 'B9',              # DIFFERENT cell!
        'critical_gaps': 'B8'                # DIFFERENT cell!
    }
}
```

**External Workbook Linking Requirements:**
For external workbook formulas to work:
1. Dashboard and source files MUST be in same directory
2. Source files MUST be named exactly:
   - ISMS-IMP-A.8.10.1.xlsx
   - ISMS-IMP-A.8.10.2.xlsx
   - ISMS-IMP-A.8.10.3.xlsx
   - ISMS-IMP-A.8.10.4.xlsx
3. Run normalize_assessment_files_a810.py to create these names
4. Excel must be configured to allow external workbook links

**Audit Considerations:**
This dashboard serves as the primary compliance artifact for ISO 27001 auditors.
Ensure:
- All source assessments are complete and approved
- External formulas display data (not #REF! errors)
- Critical gaps have remediation plans with timelines
- CISO/DPO approval is obtained before audit
- Evidence master index is comprehensive and accessible

**Data Protection:**
Dashboard contains sensitive aggregate information including:
- Overall compliance posture and critical gaps
- Risk register with information deletion risks
- Remediation plans and timelines
- Evidence inventory and locations

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Update dashboard:
- Monthly: Review KPI trends and remediation progress
- Quarterly: Update after completing quarterly assessments
- Annually: Complete dashboard refresh after annual assessments
- Ad-hoc: When critical gaps are identified or remediated

**Quality Assurance:**
Before distributing dashboard to executive stakeholders:
1. Run: python3 excel_sanity_check_a810.py (validate all workbooks)
2. Verify all external formulas display data (no #REF! errors)
3. Check compliance percentages match source assessments
4. Validate gap analysis includes all critical/high gaps
5. Confirm evidence master index is complete

**Executive Interpretation Guidance:**
For executives reviewing this dashboard:

- **Overall Compliance %:** Aggregate across all 4 domains
  - ≥90%: Strong deletion controls (green)
  - 70-89%: Adequate with gaps (yellow)
  - <70%: Significant gaps requiring attention (red)

- **Critical Gaps:** High-priority issues requiring immediate action
  - Examples: Missing retention schedules, inadequate deletion methods,
    no verification testing, shadow IT with no deletion controls

- **Risk Score:** Deletion-related risks weighted by data sensitivity
  - Score considers: Data classification, deletion method adequacy,
    verification testing, third-party controls

- **Remediation Timeline:** Prioritized action plan
  - Critical gaps: 30-60 days
  - High gaps: 60-90 days
  - Medium gaps: 90-180 days

**Regulatory Reporting:**
This dashboard supports regulatory compliance reporting for:
- ISO 27001:2022 certification audits (Clause 8.10 evidence)
- GDPR Article 5(2) accountability (demonstrating compliance)
- Swiss FADP Article 6 data processing principles
- Industry-specific audits (PCI DSS v4.0.1, HIPAA, etc.)

**Integration with Other ISMS Controls:**
A.8.10 Information Deletion integrates with:
- A.5.23: Cloud Services (third-party deletion controls)
- A.8.9: Configuration Management (deletion triggers in lifecycle)
- A.8.11: Data Masking (anonymization vs. deletion decisions)
- A.8.19: Security Logging (deletion log management)

Ensure cross-control consistency when remediating gaps.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import os
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.10.5"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.8.10"
CONTROL_NAME = "Information Deletion"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# SECTION 1: CONFIGURATION AND CONSTANTS
# ============================================================================

WORKBOOK_TITLE = "ISMS-IMP-A.8.10.5 - Compliance Dashboard (External Links)"


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
TRASH = '\U0001F5D1'  # 🗑️  Wastebasket
DISK = '\U0001F4BE'   # 💾 Floppy Disk
GLOBE = '\U0001F310'  # 🌐 Globe
SEARCH = '\U0001F50D' # 🔍 Magnifying Glass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

FILENAME_PREFIX = "ISMS-IMP-A.8.10.5_Compliance_Dashboard"
VERSION = "2.1"  # Updated with corrected cell references
RELATED_POLICY = "ISMS-POL-A.8.10 (Master), ISMS-POL-A.8.10-S4 (Governance)"

# Normalized source filenames (created by normalize_assessment_files_a810.py)
SOURCE_FILES = {
    "A.8.10.1": "ISMS-IMP-A.8.10.1.xlsx",
    "A.8.10.2": "ISMS-IMP-A.8.10.2.xlsx",
    "A.8.10.3": "ISMS-IMP-A.8.10.3.xlsx",
    "A.8.10.4": "ISMS-IMP-A.8.10.4.xlsx"
}

# VERIFIED cell locations in source assessments
# Based on actual analysis of generate_a810_X.py scripts
CELL_MAPPINGS = {
    "A.8.10.1": {
        "sheet": "Summary Dashboard",
        "compliance_pct": "G10",      # Overall compliance % in OVERALL row
        "non_compliant": "E10",       # Non-compliant count
        "total_items": "B10"          # Total items assessed
    },
    "A.8.10.2": {
        "sheet": "Summary Dashboard",
        "compliance_pct": "G10",
        "non_compliant": "E10",
        "total_items": "B10"
    },
    "A.8.10.3": {
        "sheet": "Summary Dashboard",
        "compliance_pct": "G10",
        "non_compliant": "E10",
        "total_items": "B10"
    },
    "A.8.10.4": {
        "sheet": "Verification Dashboard",  # DIFFERENT SHEET NAME!
        "compliance_pct": "B9",             # Different structure
        "critical_gaps": "B8",              # Critical gaps count
        "effectiveness": "B6"               # Average control effectiveness
    }
}

# Sheet names
SHEET_NAMES = [
    "Instructions & Legend",
    "Overall A.8.10 Compliance",
    "Retention Schedule Health",
    "Deletion Method Effectiveness",
    "Third-Party Deletion Performance",
    "Verification & Evidence Quality",
    "Critical Gaps Dashboard",
    "Trend Analysis",
    "Executive Summary"
]

# Data validation lists
VALIDATION_LISTS = {
    "rag_status": [
        "Green (On Target)",
        "Amber (Minor Issues)",
        "Red (Critical Issues)",
        "Grey (Not Assessed)"
    ],
    "trend_indicator": [
        "↑ Improved",
        f"{ARROW} Stable",
        "↓ Declined",
        "N/A"
    ],
    "compliance_status": [
        "Fully Compliant",
        "Substantially Compliant",
        "Partially Compliant",
        "Non-Compliant",
        "Not Assessed"
    ],
    "audit_readiness": [
        "Fully Ready",
        "Mostly Ready",
        "Partially Ready",
        "Not Ready"
    ],
    "gap_severity": [
        "Critical",
        "High",
        "Medium",
        "Low"
    ],
    "gap_status": [
        "Not Started",
        "Planning",
        "In Progress",
        "Testing",
        "Completed",
        "On Hold"
    ],
    "priority": [
        "P1 - Critical",
        "P2 - High",
        "P3 - Medium",
        "P4 - Low"
    ],
    "maturity_level": [
        "Optimized (90-100)",
        "Managed (80-89)",
        "Defined (70-79)",
        "Developing (60-69)",
        "Initial/Ad-Hoc (0-59)"
    ],
    "yes_no": [
        "Yes",
        "No",
        "Partial"
    ]
}

# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def create_styles():
    """Create reusable cell styles"""
    styles = {}
    
    # Header style (dark blue background, white text)
    styles['header'] = {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Subheader style (light blue background)
    styles['subheader'] = {
        'font': Font(name='Calibri', size=10, bold=True),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Input cell style (light yellow background) - for narrative sections only
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Formula cell style (light grey background) - for external links
    styles['formula_cell'] = {
        'font': Font(name='Calibri', size=10, color='0000FF'),
        'fill': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Dashboard metric cell (white background, for display)
    styles['metric_cell'] = {
        'font': Font(name='Calibri', size=10),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # Title style
    styles['title'] = {
        'font': Font(name='Calibri', size=14, bold=True, color='003366'),
        'alignment': Alignment(horizontal='left', vertical='center')
    }
    
    # Section header style
    styles['section_header'] = {
        'font': Font(name='Calibri', size=12, bold=True, color='003366'),
        'alignment': Alignment(horizontal='left', vertical='center')
    }
    
    # RAG Status colors
    styles['rag_green'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    styles['rag_amber'] = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    styles['rag_red'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    styles['rag_grey'] = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # Reference table styles
    styles['ref_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='44546A', end_color='44546A', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    styles['ref_cell'] = {
        'font': Font(name='Calibri', size=9),
        'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    return styles

# ============================================================================
# SECTION 3: HELPER FUNCTIONS
# ============================================================================

def apply_style(cell, style_dict):
    """Apply style dictionary to a cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def create_workbook():
    """Create workbook with all required sheets"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets
    for sheet_name in SHEET_NAMES:
        wb.create_sheet(title=sheet_name)
    
    return wb

# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET
# ============================================================================

def populate_instructions_sheet(ws, styles):
    """Populate Instructions & Legend sheet with external links workflow"""
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "ISMS-IMP-A.8.10.5 - Compliance Dashboard"
    apply_style(ws['A1'], styles['title'])
    ws.row_dimensions[1].height = 25
    
    # Document metadata
    ws['A3'] = "Document Information"
    ws['A3'].font = Font(bold=True, size=11)
    
    metadata = [
        ("Document ID:", "ISMS-IMP-A.8.10.5"),
        ("Title:", "Information Deletion - Compliance Dashboard"),
        ("Version:", VERSION),
        ("Related Policy:", RELATED_POLICY),
        ("ISO 27001:2022:", "Control A.8.10 (Information Deletion)"),
        ("Report Type:", "Executive Dashboard with External Links"),
        ("Last Updated:", "=TODAY()"),
        ("Review Cycle:", "Quarterly")
    ]
    
    row = 4
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "=" in str(value):
            ws[f'B{row}'].font = Font(color='0000FF')
        row += 1
    
    # Purpose section
    row += 2
    ws[f'A{row}'] = "PURPOSE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    purpose_text = """This dashboard consolidates deletion compliance metrics from the four A.8.10 assessment workbooks using EXTERNAL WORKBOOK LINKS. Unlike manual entry, this approach enables automatic data population and updates.

Key Benefits:
- Auto-populates metrics from source assessments (no manual re-entry)
- Updates automatically when source files change (click "Update Links")
- Eliminates transcription errors and data staleness
- Provides real-time compliance overview for executives
- Follows proven A.8.24 System Engineering pattern

IMPORTANT: This version uses VERIFIED cell references from actual assessment scripts."""
    
    ws.merge_cells(f'A{row}:F{row + 7}')
    cell = ws[f'A{row}']
    cell.value = purpose_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Setup instructions
    row += 9
    ws[f'A{row}'] = "⚙️ SETUP INSTRUCTIONS (MUST COMPLETE BEFORE USE)"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    setup_text = """CRITICAL: This dashboard requires normalized source files to function.

Step 1: Complete All Assessments
   • ISMS-IMP-A.8.10.1 (Retention & Deletion Triggers)
   • ISMS-IMP-A.8.10.2 (Deletion Methods)
   • ISMS-IMP-A.8.10.3 (Third-Party & Cloud Deletion)
   • ISMS-IMP-A.8.10.4 (Verification & Evidence)

Step 2: Normalize Assessment Files
   Run the normalization script:
   
   python3 normalize_assessment_files_a810.py
   
   This script:
   • Validates assessment Document IDs
   • Copies files to standardized names (removes dates/versions)
   • Creates normalized files in Dashboard_Sources folder
   • Generates audit manifest for traceability

Step 3: Place Dashboard in Correct Location
   Move THIS file (the dashboard) into the same directory as the normalized files:
   
   Dashboard_Sources/
   ├── ISMS-IMP-A.8.10.1.xlsx  (normalized)
   ├── ISMS-IMP-A.8.10.2.xlsx  (normalized)
   ├── ISMS-IMP-A.8.10.3.xlsx  (normalized)
   ├── ISMS-IMP-A.8.10.4.xlsx  (normalized)
   └── ISMS-IMP-A.8.10.5_Compliance_Dashboard_YYYYMMDD.xlsx  (THIS FILE)

Step 4: Enable External Links
   When you open this dashboard, Excel will prompt:
   
   "This workbook contains links to other data sources"
   
   Click "Update" to enable automatic data population.
   
   ⚠️ If you click "Don't Update", formulas will show #REF! errors."""
    
    ws.merge_cells(f'A{row}:F{row + 30}')
    cell = ws[f'A{row}']
    cell.value = setup_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # How it works
    row += 32
    ws[f'A{row}'] = "🔗 HOW EXTERNAL LINKS WORK"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    links_text = """The dashboard uses Excel external workbook links (formulas that reference other files).

Example Formulas (VERIFIED from actual scripts):
='[ISMS-IMP-A.8.10.1.xlsx]Summary Dashboard'!G10     (A.8.10.1 Compliance %)
='[ISMS-IMP-A.8.10.2.xlsx]Summary Dashboard'!G10     (A.8.10.2 Compliance %)
='[ISMS-IMP-A.8.10.3.xlsx]Summary Dashboard'!G10     (A.8.10.3 Compliance %)
='[ISMS-IMP-A.8.10.4.xlsx]Verification Dashboard'!B9 (A.8.10.4 Compliance %)

IMPORTANT NOTE: A.8.10.4 uses a DIFFERENT sheet name ("Verification Dashboard" 
instead of "Summary Dashboard"). This is intentional per the original script design.

This reads:
- Open file: ISMS-IMP-A.8.10.X.xlsx
- Go to sheet: Summary Dashboard (or Verification Dashboard for A.8.10.4)
- Get value from: Cell G10 (or B9 for A.8.10.4)

Cell Color Coding:
- Grey cells with blue text = External link formulas (auto-populated)
- Yellow cells = Manual entry (narrative sections only)
- White cells = Calculated from other cells

Data Flow:
A.8.10.1 Assessment ──┐
A.8.10.2 Assessment ──┼──> Normalization ──> Normalized Files ──> Dashboard ──> Auto-Populated!
A.8.10.3 Assessment ──┤    (Script)          (Standardized)       (Ext Links)
A.8.10.4 Assessment ──┘"""
    
    ws.merge_cells(f'A{row}:F{row + 19}')
    cell = ws[f'A{row}']
    cell.value = links_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Cell reference mapping
    row += 21
    ws[f'A{row}'] = "📍 VERIFIED CELL REFERENCE MAPPING"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    mapping_text = """This dashboard uses VERIFIED cell references from actual assessment generator scripts.

A.8.10.1 (Retention & Deletion Triggers):
  Sheet: "Summary Dashboard"
  Cell G10: Overall Compliance %
  Cell E10: Non-Compliant Count
  Cell B10: Total Items

A.8.10.2 (Deletion Methods):
  Sheet: "Summary Dashboard"
  Cell G10: Overall Compliance %
  Cell E10: Non-Compliant Count
  Cell B10: Total Items

A.8.10.3 (Third-Party & Cloud):
  Sheet: "Summary Dashboard"
  Cell G10: Overall Compliance %
  Cell E10: Non-Compliant Count
  Cell B10: Total Items

A.8.10.4 (Verification & Evidence):
  Sheet: "Verification Dashboard" ← DIFFERENT NAME!
  Cell B9: Compliance Status %
  Cell B8: Critical Gaps Count
  Cell B6: Average Control Effectiveness

If formulas show #REF! errors, verify these cell locations in your source assessments."""
    
    ws.merge_cells(f'A{row}:F{row + 24}')
    cell = ws[f'A{row}']
    cell.value = mapping_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Quarterly update workflow
    row += 26
    ws[f'A{row}'] = "🔄 QUARTERLY UPDATE WORKFLOW"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    update_text = """To update the dashboard with new assessment data:

1. Update source assessment files (A.8.10.1-4) with current data
2. Re-run normalization script to refresh normalized copies:
   
   python3 normalize_assessment_files_a810.py
   
3. Open this dashboard
4. Excel prompts "Update Links" → Click "Update"
5. Dashboard automatically reflects current compliance status

That's it! No manual data re-entry required.

Pro Tip: Set calendar reminder for quarterly normalization + link refresh."""
    
    ws.merge_cells(f'A{row}:F{row + 13}')
    cell = ws[f'A{row}']
    cell.value = update_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Troubleshooting
    row += 15
    ws[f'A{row}'] = "🔧 TROUBLESHOOTING"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    troubleshoot_text = """Problem: Formulas show #REF! errors
Solution: Files not in same directory. Move dashboard to Dashboard_Sources folder.

Problem: Excel doesn't prompt to update links
Solution: Go to Data → Edit Links → Update Values

Problem: Dashboard shows old data
Solution: Re-run normalization script, then click "Update Links" in Excel

Problem: "Source file not found" error
Solution: Verify normalized files exist and match expected names:
   • ISMS-IMP-A.8.10.1.xlsx
   • ISMS-IMP-A.8.10.2.xlsx
   • ISMS-IMP-A.8.10.3.xlsx
   • ISMS-IMP-A.8.10.4.xlsx

Problem: A.8.10.4 formulas show #REF! but others work
Solution: A.8.10.4 uses "Verification Dashboard" sheet name (not "Summary Dashboard").
          Verify sheet exists in ISMS-IMP-A.8.10.4.xlsx file.

Problem: Cell references seem wrong
Solution: This version uses VERIFIED references from actual scripts. Check the 
          "VERIFIED CELL REFERENCE MAPPING" section above for exact locations."""
    
    ws.merge_cells(f'A{row}:F{row + 20}')
    cell = ws[f'A{row}']
    cell.value = troubleshoot_text
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 5: SHEET 2 - OVERALL A.8.10 COMPLIANCE (CORRECTED CELL REFERENCES)
# ============================================================================

def populate_sheet2_overall(ws, styles):
    """Sheet 2: Overall A.8.10 Compliance (with CORRECTED external links)"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "OVERALL A.8.10 INFORMATION DELETION COMPLIANCE STATUS"
    apply_style(ws['A1'], styles['title'])
    ws.row_dimensions[1].height = 25
    
    # Snapshot date
    ws['A3'] = "Report Date:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = "=TODAY()"
    ws['B3'].font = Font(color='0000FF')
    
    ws['A4'] = "Reporting Period:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles['input_cell'])
    
    # Section 1: Key Compliance Metrics (External Links with CORRECT cell references)
    ws['A6'] = "SECTION 1: KEY COMPLIANCE METRICS (AUTO-POPULATED FROM SOURCE ASSESSMENTS)"
    apply_style(ws['A6'], styles['section_header'])
    ws.merge_cells('A6:E6')
    
    headers = ["Assessment Component", "Compliance Score (%)", "Non-Compliant Items", "Total Items", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # External link formulas pointing to VERIFIED cells in each assessment
    # Using the CELL_MAPPINGS defined in configuration
    assessment_data = [
        [
            "A.8.10.1: Retention & Deletion Triggers",
            f"='[{SOURCE_FILES['A.8.10.1']}]{CELL_MAPPINGS['A.8.10.1']['sheet']}'!{CELL_MAPPINGS['A.8.10.1']['compliance_pct']}",
            f"='[{SOURCE_FILES['A.8.10.1']}]{CELL_MAPPINGS['A.8.10.1']['sheet']}'!{CELL_MAPPINGS['A.8.10.1']['non_compliant']}",
            f"='[{SOURCE_FILES['A.8.10.1']}]{CELL_MAPPINGS['A.8.10.1']['sheet']}'!{CELL_MAPPINGS['A.8.10.1']['total_items']}",
            ""  # Status derived from compliance % (manual or formula)
        ],
        [
            "A.8.10.2: Deletion Methods",
            f"='[{SOURCE_FILES['A.8.10.2']}]{CELL_MAPPINGS['A.8.10.2']['sheet']}'!{CELL_MAPPINGS['A.8.10.2']['compliance_pct']}",
            f"='[{SOURCE_FILES['A.8.10.2']}]{CELL_MAPPINGS['A.8.10.2']['sheet']}'!{CELL_MAPPINGS['A.8.10.2']['non_compliant']}",
            f"='[{SOURCE_FILES['A.8.10.2']}]{CELL_MAPPINGS['A.8.10.2']['sheet']}'!{CELL_MAPPINGS['A.8.10.2']['total_items']}",
            ""
        ],
        [
            "A.8.10.3: Third-Party & Cloud Deletion",
            f"='[{SOURCE_FILES['A.8.10.3']}]{CELL_MAPPINGS['A.8.10.3']['sheet']}'!{CELL_MAPPINGS['A.8.10.3']['compliance_pct']}",
            f"='[{SOURCE_FILES['A.8.10.3']}]{CELL_MAPPINGS['A.8.10.3']['sheet']}'!{CELL_MAPPINGS['A.8.10.3']['non_compliant']}",
            f"='[{SOURCE_FILES['A.8.10.3']}]{CELL_MAPPINGS['A.8.10.3']['sheet']}'!{CELL_MAPPINGS['A.8.10.3']['total_items']}",
            ""
        ],
        [
            "A.8.10.4: Verification & Evidence",
            f"='[{SOURCE_FILES['A.8.10.4']}]{CELL_MAPPINGS['A.8.10.4']['sheet']}'!{CELL_MAPPINGS['A.8.10.4']['compliance_pct']}",
            f"='[{SOURCE_FILES['A.8.10.4']}]{CELL_MAPPINGS['A.8.10.4']['sheet']}'!{CELL_MAPPINGS['A.8.10.4']['critical_gaps']}",  # Note: using critical_gaps here
            "",  # A.8.10.4 doesn't have total_items in same structure
            ""
        ]
    ]
    
    for idx, row_data in enumerate(assessment_data, start=8):
        ws.cell(row=idx, column=1, value=row_data[0])
        for col in range(2, 6):
            cell = ws.cell(row=idx, column=col, value=row_data[col-1])
            if row_data[col-1]:  # If there's a formula
                apply_style(cell, styles['formula_cell'])
            else:  # Empty cells for manual entry
                apply_style(cell, styles['input_cell'])
    
    # Overall summary row (aggregates the above)
    ws['A12'] = "OVERALL A.8.10 COMPLIANCE"
    apply_style(ws['A12'], styles['subheader'])
    
    # Average of compliance scores (B8:B11)
    ws['B12'] = "=IFERROR(AVERAGE(B8:B11),0)"
    apply_style(ws['B12'], styles['formula_cell'])
    
    # Sum of non-compliant items (C8:C10, excluding C11 which is critical gaps from A.8.10.4)
    ws['C12'] = "=IFERROR(SUM(C8:C10)+C11,0)"
    apply_style(ws['C12'], styles['formula_cell'])
    
    # Sum of total items (D8:D10, D11 is empty for A.8.10.4)
    ws['D12'] = "=IFERROR(SUM(D8:D10),0)"
    apply_style(ws['D12'], styles['formula_cell'])
    
    # Overall status based on average compliance
    ws['E12'] = "=IF(B12>=95,\"Fully Compliant\",IF(B12>=80,\"Substantially Compliant\",IF(B12>=60,\"Partially Compliant\",IF(B12>0,\"Non-Compliant\",\"Not Assessed\"))))"
    apply_style(ws['E12'], styles['formula_cell'])
    
    # Section 2: Component Status Details
    ws['A14'] = "SECTION 2: COMPONENT STATUS DETAILS"
    apply_style(ws['A14'], styles['section_header'])
    ws.merge_cells('A14:D14')
    
    component_headers = ["Component", "Key Strength", "Primary Gap", "Priority Action"]
    for col_idx, header in enumerate(component_headers, start=1):
        cell = ws.cell(row=15, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # Component details (manual entry - extracted from source assessments)
    component_names = [
        "A.8.10.1: Retention Schedules",
        "A.8.10.2: Deletion Methods",
        "A.8.10.3: Third-Party Deletion",
        "A.8.10.4: Verification & Evidence"
    ]
    
    for idx, name in enumerate(component_names, start=16):
        ws.cell(row=idx, column=1, value=name)
        for col in range(2, 5):
            apply_style(ws.cell(row=idx, column=col), styles['input_cell'])
    
    # Section 3: Critical Findings Requiring Immediate Attention
    ws['A21'] = "SECTION 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    apply_style(ws['A21'], styles['section_header'])
    ws.merge_cells('A21:F21')
    
    findings_headers = ["Finding", "Source Assessment", "Severity", "Impact", "Owner", "Due Date"]
    for col_idx, header in enumerate(findings_headers, start=1):
        cell = ws.cell(row=22, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # Top 10 findings rows (manual entry - consolidated from all 4 assessments)
    for row in range(23, 33):
        for col in range(1, 7):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Data validation for severity
    dv_severity = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["gap_severity"])}"', allow_blank=True)
    ws.add_data_validation(dv_severity)
    dv_severity.add('C23:C32')
    
    # Section 4: Compliance Trend Indicators
    ws['A34'] = "SECTION 4: COMPLIANCE TREND (vs. Previous Quarter)"
    apply_style(ws['A34'], styles['section_header'])
    ws.merge_cells('A34:D34')
    
    trend_headers = ["Metric", "Current", "Previous", "Trend"]
    for col_idx, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=35, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    trend_metrics = [
        ["Overall Compliance %", "=B12", "", ""],
        ["Total Non-Compliant Items", "=C12", "", ""],
        ["Critical Gaps", "=C11", "", ""],  # From A.8.10.4
        ["A.8.10.1 Compliance %", "=B8", "", ""],
        ["A.8.10.2 Compliance %", "=B9", "", ""],
        ["A.8.10.3 Compliance %", "=B10", "", ""],
        ["A.8.10.4 Compliance %", "=B11", "", ""]
    ]
    
    for idx, metric in enumerate(trend_metrics, start=36):
        ws.cell(row=idx, column=1, value=metric[0])
        
        # Current value (formula reference)
        cell_current = ws.cell(row=idx, column=2, value=metric[1])
        if "=" in metric[1]:
            apply_style(cell_current, styles['formula_cell'])
        else:
            apply_style(cell_current, styles['input_cell'])
        
        # Previous value (manual entry from last quarter)
        apply_style(ws.cell(row=idx, column=3), styles['input_cell'])
        
        # Trend indicator (dropdown)
        apply_style(ws.cell(row=idx, column=4), styles['input_cell'])
    
    # Data validation for trend indicators
    dv_trend = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["trend_indicator"])}"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add('D36:D42')
    
    # Notes section
    ws['A44'] = "NOTES:"
    ws['A44'].font = Font(bold=True)
    ws.merge_cells('A45:F48')
    cell_notes = ws['A45']
    apply_style(cell_notes, styles['input_cell'])
    cell_notes.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 6: SHEET 3 - RETENTION SCHEDULE HEALTH (A.8.10.1 DATA)
# ============================================================================

def populate_sheet3_retention(ws, styles):
    """Sheet 3: Retention Schedule Health (with external links from A.8.10.1)"""
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "RETENTION SCHEDULE HEALTH (A.8.10.1 METRICS)"
    apply_style(ws['A1'], styles['title'])
    
    # Key metrics from A.8.10.1 (external links to VERIFIED cells)
    ws['A3'] = "KEY RETENTION METRICS (AUTO-POPULATED FROM A.8.10.1)"
    apply_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')
    
    metric_headers = ["Metric", "Current Value", "Target", "Status"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # External links to A.8.10.1 Summary Dashboard (VERIFIED cells)
    retention_metrics = [
        [
            "Overall A.8.10.1 Compliance %",
            f"='[{SOURCE_FILES['A.8.10.1']}]{CELL_MAPPINGS['A.8.10.1']['sheet']}'!{CELL_MAPPINGS['A.8.10.1']['compliance_pct']}",
            "≥95%",
            ""
        ],
        [
            "Non-Compliant Items",
            f"='[{SOURCE_FILES['A.8.10.1']}]{CELL_MAPPINGS['A.8.10.1']['sheet']}'!{CELL_MAPPINGS['A.8.10.1']['non_compliant']}",
            "0",
            ""
        ],
        [
            "Total Items Assessed",
            f"='[{SOURCE_FILES['A.8.10.1']}]{CELL_MAPPINGS['A.8.10.1']['sheet']}'!{CELL_MAPPINGS['A.8.10.1']['total_items']}",
            "All Categories",
            ""
        ]
    ]
    
    for idx, row_data in enumerate(retention_metrics, start=5):
        ws.cell(row=idx, column=1, value=row_data[0])
        
        # Current value (external link formula)
        cell_b = ws.cell(row=idx, column=2, value=row_data[1])
        apply_style(cell_b, styles['formula_cell'])
        
        # Target
        ws.cell(row=idx, column=3, value=row_data[2])
        
        # Status (manual entry or dropdown)
        apply_style(ws.cell(row=idx, column=4), styles['input_cell'])
    
    # Data validation for status
    dv_status = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["rag_status"])}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('D5:D7')
    
    # Detailed metrics (manual entry - source assessment has dynamic row structure)
    ws['A9'] = "DETAILED RETENTION SCHEDULE METRICS (MANUAL ENTRY)"
    apply_style(ws['A9'], styles['section_header'])
    ws.merge_cells('A9:D9')
    
    ws['A10'] = "Note: These metrics have dynamic row locations in source assessment."
    ws['A10'].font = Font(italic=True, size=9)
    ws.merge_cells('A10:D10')
    
    detailed_headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(detailed_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    detailed_metrics = [
        "Total Data Categories Identified",
        "Data Categories with Defined Retention",
        "Data Categories with Legal Basis",
        "Event-Based Retention Categories",
        "Categories with PII/SPI",
        "Categories Pending Review",
        "Categories with Automated Triggers",
        "Categories with Manual Triggers Only",
        "Legal Hold Breaches (Last 12 Months)",
        "Average Data Subject Request Response Time (Days)"
    ]
    
    for idx, metric in enumerate(detailed_metrics, start=12):
        ws.cell(row=idx, column=1, value=metric)
        for col in range(2, 5):
            apply_style(ws.cell(row=idx, column=col), styles['input_cell'])
    
    # Gap analysis
    ws['A23'] = "RETENTION SCHEDULE GAPS"
    apply_style(ws['A23'], styles['section_header'])
    ws.merge_cells('A23:E23')
    
    gap_headers = ["Data Category", "Current Gap", "Risk Level", "Remediation Plan", "Target Date"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=24, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    for row in range(25, 35):  # 10 gap rows
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 7: SHEET 4 - DELETION METHOD EFFECTIVENESS (A.8.10.2 DATA)
# ============================================================================

def populate_sheet4_methods(ws, styles):
    """Sheet 4: Deletion Method Effectiveness (with external links from A.8.10.2)"""
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "DELETION METHOD EFFECTIVENESS (A.8.10.2 METRICS)"
    apply_style(ws['A1'], styles['title'])
    
    # Key metrics from A.8.10.2 (external links to VERIFIED cells)
    ws['A3'] = "DELETION METHOD COMPLIANCE (AUTO-POPULATED FROM A.8.10.2)"
    apply_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')
    
    metric_headers = ["Metric", "Current Value", "Target", "Status"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # External links to A.8.10.2 Summary Dashboard (VERIFIED cells)
    deletion_metrics = [
        [
            "Overall A.8.10.2 Compliance %",
            f"='[{SOURCE_FILES['A.8.10.2']}]{CELL_MAPPINGS['A.8.10.2']['sheet']}'!{CELL_MAPPINGS['A.8.10.2']['compliance_pct']}",
            "≥95%",
            ""
        ],
        [
            "Non-Compliant Items",
            f"='[{SOURCE_FILES['A.8.10.2']}]{CELL_MAPPINGS['A.8.10.2']['sheet']}'!{CELL_MAPPINGS['A.8.10.2']['non_compliant']}",
            "0",
            ""
        ],
        [
            "Total Media Types Assessed",
            f"='[{SOURCE_FILES['A.8.10.2']}]{CELL_MAPPINGS['A.8.10.2']['sheet']}'!{CELL_MAPPINGS['A.8.10.2']['total_items']}",
            "All Media Types",
            ""
        ]
    ]
    
    for idx, row_data in enumerate(deletion_metrics, start=5):
        ws.cell(row=idx, column=1, value=row_data[0])
        
        # Current value (external link formula)
        cell_b = ws.cell(row=idx, column=2, value=row_data[1])
        apply_style(cell_b, styles['formula_cell'])
        
        # Target
        ws.cell(row=idx, column=3, value=row_data[2])
        
        # Status (manual entry or dropdown)
        apply_style(ws.cell(row=idx, column=4), styles['input_cell'])
    
    # Data validation for status
    dv_status = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["rag_status"])}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('D5:D7')
    
    # Detailed metrics by media type (manual entry)
    ws['A9'] = "DELETION METHOD METRICS BY MEDIA TYPE (MANUAL ENTRY)"
    apply_style(ws['A9'], styles['section_header'])
    ws.merge_cells('A9:D9')
    
    ws['A10'] = "Note: Extract these from A.8.10.2 assessment sheets (dynamic row locations)."
    ws['A10'].font = Font(italic=True, size=9)
    ws.merge_cells('A10:D10')
    
    detailed_headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(detailed_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    detailed_metrics = [
        "NIST SP 800-88 Rev. 2 Compliant Methods (%)",
        "SSD Secure Erase Available",
        "HDD Degaussing Capability",
        "Cloud Storage Deletion SLA Documented",
        "Database Deletion Methods Tested",
        "Mobile Device Secure Wipe Available",
        "Deletion Methods Verified (Testing %)",
        "Physical Media Destruction Process",
        "Backup Deletion Coordination",
        "SSD Overwrite Detection Implemented"
    ]
    
    for idx, metric in enumerate(detailed_metrics, start=12):
        ws.cell(row=idx, column=1, value=metric)
        for col in range(2, 5):
            apply_style(ws.cell(row=idx, column=col), styles['input_cell'])
    
    # Method-specific gaps
    ws['A23'] = "DELETION METHOD GAPS BY MEDIA TYPE"
    apply_style(ws['A23'], styles['section_header'])
    ws.merge_cells('A23:E23')
    
    method_headers = ["Media Type", "Current Method", "Compliance Gap", "Recommended Method", "Implementation Status"]
    for col_idx, header in enumerate(method_headers, start=1):
        cell = ws.cell(row=24, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    for row in range(25, 35):  # 10 rows
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    
    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 8: SHEET 5 - THIRD-PARTY DELETION PERFORMANCE (A.8.10.3 DATA)
# ============================================================================

def populate_sheet5_thirdparty(ws, styles):
    """Sheet 5: Third-Party Deletion Performance (with external links from A.8.10.3)"""
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "THIRD-PARTY & CLOUD DELETION PERFORMANCE (A.8.10.3 METRICS)"
    apply_style(ws['A1'], styles['title'])
    
    # Key metrics from A.8.10.3 (external links to VERIFIED cells)
    ws['A3'] = "THIRD-PARTY DELETION METRICS (AUTO-POPULATED FROM A.8.10.3)"
    apply_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')
    
    metric_headers = ["Metric", "Current Value", "Target", "Status"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # External links to A.8.10.3 Summary Dashboard (VERIFIED cells)
    thirdparty_metrics = [
        [
            "Overall A.8.10.3 Compliance %",
            f"='[{SOURCE_FILES['A.8.10.3']}]{CELL_MAPPINGS['A.8.10.3']['sheet']}'!{CELL_MAPPINGS['A.8.10.3']['compliance_pct']}",
            "≥95%",
            ""
        ],
        [
            "Non-Compliant Items",
            f"='[{SOURCE_FILES['A.8.10.3']}]{CELL_MAPPINGS['A.8.10.3']['sheet']}'!{CELL_MAPPINGS['A.8.10.3']['non_compliant']}",
            "0",
            ""
        ],
        [
            "Total Providers Assessed",
            f"='[{SOURCE_FILES['A.8.10.3']}]{CELL_MAPPINGS['A.8.10.3']['sheet']}'!{CELL_MAPPINGS['A.8.10.3']['total_items']}",
            "All Providers",
            ""
        ]
    ]
    
    for idx, row_data in enumerate(thirdparty_metrics, start=5):
        ws.cell(row=idx, column=1, value=row_data[0])
        
        # Current value (external link formula)
        cell_b = ws.cell(row=idx, column=2, value=row_data[1])
        apply_style(cell_b, styles['formula_cell'])
        
        # Target
        ws.cell(row=idx, column=3, value=row_data[2])
        
        # Status (manual entry or dropdown)
        apply_style(ws.cell(row=idx, column=4), styles['input_cell'])
    
    # Data validation for status
    dv_status = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["rag_status"])}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('D5:D7')
    
    # Detailed third-party metrics (manual entry)
    ws['A9'] = "DETAILED THIRD-PARTY METRICS (MANUAL ENTRY)"
    apply_style(ws['A9'], styles['section_header'])
    ws.merge_cells('A9:D9')
    
    ws['A10'] = "Note: Extract these from A.8.10.3 assessment sheets (dynamic row locations)."
    ws['A10'].font = Font(italic=True, size=9)
    ws.merge_cells('A10:D10')
    
    detailed_headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(detailed_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    detailed_metrics = [
        "Deletion Clauses in Contracts (%)",
        "Deletion Certificates Received (%)",
        "Average Certificate Quality Score (0-10)",
        "SLA Compliance Rate (%)",
        "GDPR Article 28 Compliance (%)",
        "Subprocessors Mapped (%)",
        "Shadow IT Systems Detected",
        "Cloud Provider API Deletion Tested",
        "Vendor Audit Rights Documented",
        "Data Residency Requirements Met"
    ]
    
    for idx, metric in enumerate(detailed_metrics, start=12):
        ws.cell(row=idx, column=1, value=metric)
        for col in range(2, 5):
            apply_style(ws.cell(row=idx, column=col), styles['input_cell'])
    
    # High-risk providers
    ws['A23'] = "HIGH-RISK THIRD-PARTY PROVIDERS"
    apply_style(ws['A23'], styles['section_header'])
    ws.merge_cells('A23:G23')
    
    provider_headers = ["Provider Name", "Service Type", "Risk Issue", "Certificate Quality", "SLA Status", "Action Required", "Owner"]
    for col_idx, header in enumerate(provider_headers, start=1):
        cell = ws.cell(row=24, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    for row in range(25, 35):  # 10 rows
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    
    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 9: SHEET 6 - VERIFICATION & EVIDENCE QUALITY (A.8.10.4 DATA)
# ============================================================================

def populate_sheet6_verification(ws, styles):
    """Sheet 6: Verification & Evidence Quality (with external links from A.8.10.4)"""
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "VERIFICATION & EVIDENCE QUALITY (A.8.10.4 METRICS)"
    apply_style(ws['A1'], styles['title'])
    
    # Key metrics from A.8.10.4 (external links to VERIFIED cells)
    ws['A3'] = "VERIFICATION & EVIDENCE METRICS (AUTO-POPULATED FROM A.8.10.4)"
    apply_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')
    
    metric_headers = ["Metric", "Current Value", "Target", "Status"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # External links to A.8.10.4 Verification Dashboard (VERIFIED cells - DIFFERENT STRUCTURE!)
    verification_metrics = [
        [
            "Overall A.8.10.4 Compliance %",
            f"='[{SOURCE_FILES['A.8.10.4']}]{CELL_MAPPINGS['A.8.10.4']['sheet']}'!{CELL_MAPPINGS['A.8.10.4']['compliance_pct']}",
            "≥95%",
            ""
        ],
        [
            "Critical Gaps Count",
            f"='[{SOURCE_FILES['A.8.10.4']}]{CELL_MAPPINGS['A.8.10.4']['sheet']}'!{CELL_MAPPINGS['A.8.10.4']['critical_gaps']}",
            "0",
            ""
        ],
        [
            "Average Control Effectiveness",
            f"='[{SOURCE_FILES['A.8.10.4']}]{CELL_MAPPINGS['A.8.10.4']['sheet']}'!{CELL_MAPPINGS['A.8.10.4']['effectiveness']}",
            "Effective or Higher",
            ""
        ]
    ]
    
    for idx, row_data in enumerate(verification_metrics, start=5):
        ws.cell(row=idx, column=1, value=row_data[0])
        
        # Current value (external link formula)
        cell_b = ws.cell(row=idx, column=2, value=row_data[1])
        apply_style(cell_b, styles['formula_cell'])
        
        # Target
        ws.cell(row=idx, column=3, value=row_data[2])
        
        # Status (manual entry or dropdown)
        apply_style(ws.cell(row=idx, column=4), styles['input_cell'])
    
    # Data validation for status
    dv_status = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["rag_status"])}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('D5:D7')
    
    # Detailed verification metrics (manual entry)
    ws['A9'] = "DETAILED VERIFICATION METRICS (MANUAL ENTRY)"
    apply_style(ws['A9'], styles['section_header'])
    ws.merge_cells('A9:D9')
    
    ws['A10'] = "Note: Extract these from A.8.10.4 assessment sheets (dynamic row locations)."
    ws['A10'].font = Font(italic=True, size=9)
    ws.merge_cells('A10:D10')
    
    detailed_headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(detailed_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    detailed_metrics = [
        "Deletion Events with Evidence (%)",
        "Log Retention Compliance (%)",
        "Audit Trail Completeness (%)",
        "Average Evidence Quality Score (0-10)",
        "Re-verification Test Pass Rate (%)",
        "Evidence Accessibility (24hrs) (%)",
        "Certificate Management Process",
        "Deletion Logging Coverage (%)",
        "Evidence Repository Capacity",
        "Verification Testing Frequency"
    ]
    
    for idx, metric in enumerate(detailed_metrics, start=12):
        ws.cell(row=idx, column=1, value=metric)
        for col in range(2, 5):
            apply_style(ws.cell(row=idx, column=col), styles['input_cell'])
    
    # Evidence quality gaps
    ws['A23'] = "EVIDENCE QUALITY GAPS"
    apply_style(ws['A23'], styles['section_header'])
    ws.merge_cells('A23:F23')
    
    evidence_headers = ["Evidence Type", "Current Gap", "Risk Level", "Quality Score", "Remediation Plan", "Owner"]
    for col_idx, header in enumerate(evidence_headers, start=1):
        cell = ws.cell(row=24, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    for row in range(25, 35):  # 10 rows
        for col in range(1, 7):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 20
    
    ws.freeze_panes = 'A5'

# ============================================================================
# SECTION 10: SHEET 7 - CRITICAL GAPS DASHBOARD
# ============================================================================

def populate_sheet7_gaps(ws, styles):
    """Sheet 7: Critical Gaps Dashboard (Top 20)"""
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "CRITICAL GAPS DASHBOARD - TOP 20 PRIORITIES"
    apply_style(ws['A1'], styles['title'])
    
    ws['A3'] = "Instructions: List the top 20 critical gaps across all A.8.10 assessments"
    ws['A3'].font = Font(italic=True)
    ws.merge_cells('A3:H3')
    
    ws['A4'] = "Source: Extract from each assessment's gap analysis sections"
    ws['A4'].font = Font(italic=True, size=9)
    ws.merge_cells('A4:H4')
    
    # Headers
    gap_headers = ["#", "Gap Description", "Source", "Severity", "Impact", "Owner", "Due Date", "Status"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # 20 gap rows
    for row in range(6, 26):
        ws.cell(row=row, column=1, value=row-5)  # Gap number
        for col in range(2, 9):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Data validations
    dv_severity = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["gap_severity"])}"', allow_blank=True)
    ws.add_data_validation(dv_severity)
    dv_severity.add('D6:D25')
    
    dv_status = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["gap_status"])}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('H6:H25')
    
    # Summary section
    ws['A27'] = "GAP SUMMARY"
    apply_style(ws['A27'], styles['section_header'])
    ws.merge_cells('A27:C27')
    
    ws['A28'] = "Critical Gaps:"
    ws['A28'].font = Font(bold=True)
    ws['B28'] = '=COUNTIF(D6:D25,"Critical")'
    apply_style(ws['B28'], styles['formula_cell'])
    
    ws['A29'] = "High Priority Gaps:"
    ws['A29'].font = Font(bold=True)
    ws['B29'] = '=COUNTIF(D6:D25,"High")'
    apply_style(ws['B29'], styles['formula_cell'])
    
    ws['A30'] = "Medium Priority Gaps:"
    ws['A30'].font = Font(bold=True)
    ws['B30'] = '=COUNTIF(D6:D25,"Medium")'
    apply_style(ws['B30'], styles['formula_cell'])
    
    ws['A31'] = "Low Priority Gaps:"
    ws['A31'].font = Font(bold=True)
    ws['B31'] = '=COUNTIF(D6:D25,"Low")'
    apply_style(ws['B31'], styles['formula_cell'])
    
    ws['A33'] = "Completed:"
    ws['A33'].font = Font(bold=True)
    ws['B33'] = '=COUNTIF(H6:H25,"Completed")'
    apply_style(ws['B33'], styles['formula_cell'])
    
    ws['A34'] = "In Progress:"
    ws['A34'].font = Font(bold=True)
    ws['B34'] = '=COUNTIF(H6:H25,"In Progress")'
    apply_style(ws['B34'], styles['formula_cell'])
    
    ws['A35'] = "Not Started:"
    ws['A35'].font = Font(bold=True)
    ws['B35'] = '=COUNTIF(H6:H25,"Not Started")'
    apply_style(ws['B35'], styles['formula_cell'])
    
    ws['A37'] = "Overall Completion Rate:"
    ws['A37'].font = Font(bold=True)
    ws['B37'] = '=IFERROR(B33/COUNTA(H6:H25)*100&"%",0)'
    apply_style(ws['B37'], styles['formula_cell'])
    
    # By source breakdown
    ws['D27'] = "BY SOURCE ASSESSMENT"
    apply_style(ws['D27'], styles['section_header'])
    ws.merge_cells('D27:E27')
    
    ws['D28'] = "A.8.10.1 Gaps:"
    ws['D28'].font = Font(bold=True)
    ws['E28'] = '=COUNTIF(C6:C25,"*8.10.1*")'
    apply_style(ws['E28'], styles['formula_cell'])
    
    ws['D29'] = "A.8.10.2 Gaps:"
    ws['D29'].font = Font(bold=True)
    ws['E29'] = '=COUNTIF(C6:C25,"*8.10.2*")'
    apply_style(ws['E29'], styles['formula_cell'])
    
    ws['D30'] = "A.8.10.3 Gaps:"
    ws['D30'].font = Font(bold=True)
    ws['E30'] = '=COUNTIF(C6:C25,"*8.10.3*")'
    apply_style(ws['E30'], styles['formula_cell'])
    
    ws['D31'] = "A.8.10.4 Gaps:"
    ws['D31'].font = Font(bold=True)
    ws['E31'] = '=COUNTIF(C6:C25,"*8.10.4*")'
    apply_style(ws['E31'], styles['formula_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    ws.freeze_panes = 'A6'

# ============================================================================
# SECTION 11: SHEET 8 - TREND ANALYSIS
# ============================================================================

def populate_sheet8_trends(ws, styles):
    """Sheet 8: Trend Analysis (Quarterly Snapshots)"""
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "TREND ANALYSIS - QUARTERLY COMPLIANCE SNAPSHOTS"
    apply_style(ws['A1'], styles['title'])
    
    ws['A3'] = "Instructions: Capture quarterly snapshot metrics to track improvement trends"
    ws['A3'].font = Font(italic=True)
    ws.merge_cells('A3:H3')
    
    ws['A4'] = "Note: First quarter auto-populates from current data. Subsequent quarters are manual entry."
    ws['A4'].font = Font(italic=True, size=9)
    ws.merge_cells('A4:H4')
    
    # Headers
    trend_headers = ["Metric", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Q1 2026", "Q2 2026", "Trend"]
    for col_idx, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # Metrics to track (first quarter pulls from Sheet 2, rest are manual)
    metrics = [
        ["Overall Compliance Score (%)", "='Overall A.8.10 Compliance'!B12", "", "", "", "", "", ""],
        ["Critical Gaps (Count)", "='Critical Gaps Dashboard'!B28", "", "", "", "", "", ""],
        ["Total Non-Compliant Items", "='Overall A.8.10 Compliance'!C12", "", "", "", "", "", ""],
        ["A.8.10.1 Compliance %", "='Overall A.8.10 Compliance'!B8", "", "", "", "", "", ""],
        ["A.8.10.2 Compliance %", "='Overall A.8.10 Compliance'!B9", "", "", "", "", "", ""],
        ["A.8.10.3 Compliance %", "='Overall A.8.10 Compliance'!B10", "", "", "", "", "", ""],
        ["A.8.10.4 Compliance %", "='Overall A.8.10 Compliance'!B11", "", "", "", "", "", ""],
        ["Retention Schedule Coverage (%)", "", "", "", "", "", "", ""],
        ["Deletion Method NIST Compliance (%)", "", "", "", "", "", "", ""],
        ["Third-Party Certificate Quality (Avg)", "", "", "", "", "", "", ""],
        ["Evidence Quality Score (Avg)", "", "", "", "", "", "", ""],
        ["GDPR Article 17 Readiness (%)", "", "", "", "", "", "", ""],
        ["Automated Deletion Rate (%)", "", "", "", "", "", "", ""],
        ["Audit Readiness Score (%)", "", "", "", "", "", "", ""]
    ]
    
    for idx, metric in enumerate(metrics, start=6):
        ws.cell(row=idx, column=1, value=metric[0])
        
        # Q1 column (B) - formula reference to current data
        cell_q1 = ws.cell(row=idx, column=2, value=metric[1])
        if "=" in metric[1]:
            apply_style(cell_q1, styles['formula_cell'])
        else:
            apply_style(cell_q1, styles['input_cell'])
        
        # Q2-Q2 2026 (columns C-G) - manual entry
        for col in range(3, 8):
            apply_style(ws.cell(row=idx, column=col), styles['input_cell'])
        
        # Trend indicator (H) - dropdown
        apply_style(ws.cell(row=idx, column=8), styles['input_cell'])
    
    # Data validation for trend
    dv_trend = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["trend_indicator"])}"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add('H6:H19')
    
    # Notes section
    ws['A21'] = "QUARTERLY NOTES & OBSERVATIONS"
    apply_style(ws['A21'], styles['section_header'])
    ws.merge_cells('A21:H21')
    
    quarter_notes = ["Q1 2025 Notes:", "Q2 2025 Notes:", "Q3 2025 Notes:", "Q4 2025 Notes:", "Q1 2026 Notes:", "Q2 2026 Notes:"]
    for idx, note in enumerate(quarter_notes, start=22):
        ws.cell(row=idx, column=1, value=note)
        ws.cell(row=idx, column=1).font = Font(bold=True)
        ws.merge_cells(f'B{idx}:H{idx}')
        apply_style(ws.cell(row=idx, column=2), styles['input_cell'])
        ws.cell(row=idx, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[idx].height = 40
    
    # Key milestones
    ws['A29'] = "KEY MILESTONES & EVENTS"
    apply_style(ws['A29'], styles['section_header'])
    ws.merge_cells('A29:H29')
    
    milestone_headers = ["Date", "Milestone/Event", "Impact on Compliance", "Related Quarter"]
    for col_idx, header in enumerate(milestone_headers, start=1):
        cell = ws.cell(row=30, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    for row in range(31, 41):  # 10 milestone rows
        for col in range(1, 5):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['H'].width = 15
    
    ws.freeze_panes = 'A6'

# ============================================================================
# SECTION 12: SHEET 9 - EXECUTIVE SUMMARY
# ============================================================================

def populate_sheet9_executive(ws, styles):
    """Sheet 9: Executive Summary"""
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "EXECUTIVE SUMMARY - A.8.10 INFORMATION DELETION PROGRAM"
    apply_style(ws['A1'], styles['title'])
    ws.row_dimensions[1].height = 30
    
    # Overall status
    ws['A3'] = "OVERALL DELETION PROGRAM STATUS"
    apply_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')
    
    status_headers = ["Category", "Score", "Target", "Status"]
    for col_idx, header in enumerate(status_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # These pull from Sheet 2 (Overall Compliance)
    status_data = [
        ["Overall A.8.10 Compliance", "='Overall A.8.10 Compliance'!B12", "≥95%", "='Overall A.8.10 Compliance'!E12"],
        ["Critical Gaps", "='Critical Gaps Dashboard'!B28", "0", ""],
        ["Total Non-Compliant Items", "='Overall A.8.10 Compliance'!C12", "≤5", ""],
        ["Audit Readiness", "", "Fully Ready", ""]
    ]
    
    for idx, row_data in enumerate(status_data, start=5):
        ws.cell(row=idx, column=1, value=row_data[0])
        
        # Score column
        cell_b = ws.cell(row=idx, column=2, value=row_data[1])
        if "=" in row_data[1]:
            apply_style(cell_b, styles['formula_cell'])
        else:
            apply_style(cell_b, styles['input_cell'])
        
        # Target column
        ws.cell(row=idx, column=3, value=row_data[2])
        
        # Status column
        cell_d = ws.cell(row=idx, column=4, value=row_data[3])
        if "=" in row_data[3]:
            apply_style(cell_d, styles['formula_cell'])
        else:
            apply_style(cell_d, styles['input_cell'])
    
    # Data validation for audit readiness
    dv_audit = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["audit_readiness"])}"', allow_blank=True)
    ws.add_data_validation(dv_audit)
    dv_audit.add('D8')
    
    # Executive summary text
    ws['A11'] = "EXECUTIVE SUMMARY TEXT:"
    ws['A11'].font = Font(bold=True, size=10)
    
    ws['A12'] = "Provide a brief narrative summary for executive leadership (max 500 words)."
    ws['A12'].font = Font(italic=True, size=9)
    
    for row in range(13, 28):  # 15 rows for text
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1)
        apply_style(cell, styles['input_cell'])
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Maturity scoring
    ws['A30'] = "SECTION 2: OVERALL DELETION PROGRAM MATURITY SCORING"
    apply_style(ws['A30'], styles['section_header'])
    ws.merge_cells('A30:D30')
    
    maturity_headers = ["Component", "Weight", "Score (0-100)", "Weighted Score"]
    for col_idx, header in enumerate(maturity_headers, start=1):
        cell = ws.cell(row=31, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    # Pull scores from individual component sheets (Sheets 3-6)
    maturity_data = [
        ["Retention Schedule Management (A.8.10.1)", "25%", "='Retention Schedule Health'!B5", "=IF(ISNUMBER(C32),C32*0.25,0)"],
        ["Deletion Method Compliance (A.8.10.2)", "30%", "='Deletion Method Effectiveness'!B5", "=IF(ISNUMBER(C33),C33*0.30,0)"],
        ["Third-Party Coordination (A.8.10.3)", "25%", "='Third-Party Deletion Performance'!B5", "=IF(ISNUMBER(C34),C34*0.25,0)"],
        ["Verification & Evidence (A.8.10.4)", "20%", "='Verification & Evidence Quality'!B5", "=IF(ISNUMBER(C35),C35*0.20,0)"],
        ["OVERALL A.8.10 MATURITY SCORE", "100%", "=IFERROR(AVERAGE(C32:C35),0)", "=IFERROR(SUM(D32:D35),0)"]
    ]
    
    for idx, row_data in enumerate(maturity_data, start=32):
        ws.cell(row=idx, column=1, value=row_data[0])
        ws.cell(row=idx, column=2, value=row_data[1])
        
        # Score and Weighted Score - formulas
        ws.cell(row=idx, column=3, value=row_data[2])
        apply_style(ws.cell(row=idx, column=3), styles['formula_cell'])
        
        ws.cell(row=idx, column=4, value=row_data[3])
        apply_style(ws.cell(row=idx, column=4), styles['formula_cell'])
        
        if idx == 36:  # Overall row
            apply_style(ws.cell(row=idx, column=1), styles['subheader'])
            apply_style(ws.cell(row=idx, column=2), styles['subheader'])
    
    # Maturity interpretation table
    ws['A38'] = "MATURITY LEVEL INTERPRETATION"
    ws['A38'].font = Font(bold=True, size=10)
    
    interpretation_data = [
        ["Score Range", "Maturity Level", "Description"],
        ["90-100", "Optimized", "Continuous improvement, industry leading"],
        ["80-89", "Managed", "Effective controls, minor improvements needed"],
        ["70-79", "Defined", "Basic compliance, significant gaps remain"],
        ["60-69", "Developing", "Major gaps, regulatory risk exposure"],
        ["0-59", "Initial/Ad-Hoc", "Critical gaps, immediate action required"]
    ]
    
    for idx, row_data in enumerate(interpretation_data, start=39):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            if idx == 39:
                apply_style(cell, styles['ref_header'])
            else:
                apply_style(cell, styles['ref_cell'])
    
    # Current maturity level display
    ws['A46'] = "Current Maturity Level:"
    ws['A46'].font = Font(bold=True, size=11)
    ws['B46'] = '=IF(D36>=90,"Optimized (90-100)",IF(D36>=80,"Managed (80-89)",IF(D36>=70,"Defined (70-79)",IF(D36>=60,"Developing (60-69)",IF(D36>0,"Initial/Ad-Hoc (0-59)","Not Assessed")))))'
    apply_style(ws['B46'], styles['formula_cell'])
    ws['B46'].font = Font(bold=True, size=11, color='0000FF')
    
    # Recommendations
    ws['A48'] = "SECTION 3: RECOMMENDATIONS & ACTION PLAN"
    apply_style(ws['A48'], styles['section_header'])
    ws.merge_cells('A48:E48')
    
    rec_headers = ["Recommendation", "Priority", "Owner", "Target Date", "Budget/Resources"]
    for col_idx, header in enumerate(rec_headers, start=1):
        cell = ws.cell(row=49, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    for row in range(50, 55):  # 5 recommendation rows
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Data validation for priority
    dv_priority = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LISTS["priority"])}"', allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add('B50:B54')
    
    # Approval section
    ws['A57'] = "SECTION 4: EXECUTIVE APPROVAL"
    apply_style(ws['A57'], styles['section_header'])
    ws.merge_cells('A57:D57')
    
    approval_headers = ["Role", "Name", "Signature", "Date"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=58, column=col_idx, value=header)
        apply_style(cell, styles['header'])
    
    approval_data = [
        ["Assessor", "", "", ""],
        ["Compliance Manager", "", "", ""],
        ["CISO", "", "", ""],
        ["DPO", "", "", ""]
    ]
    
    for idx, row_data in enumerate(approval_data, start=59):
        ws.cell(row=idx, column=1, value=row_data[0])
        for col in range(2, 5):
            apply_style(ws.cell(row=idx, column=col), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    
    ws.freeze_panes = 'A12'

# ============================================================================
# SECTION 13: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.8.10.5 - Compliance Dashboard")
        logger.info("(EXTERNAL LINKS VERSION - CORRECTED CELL REFERENCES)")
        logger.info("=" * 70)
        logger.info("IMPORTANT: This dashboard uses EXTERNAL WORKBOOK LINKS")
        logger.info("  with VERIFIED cell references from actual A.8.10 scripts")
        logger.info("  Prerequisites: Complete A.8.10.1-4 assessments, run normalize script")

        # Create workbook
        logger.info("Creating workbook structure...")
        wb = create_workbook()
        styles = create_styles()

        # Populate each sheet
        logger.info("Populating Instructions sheet...")
        populate_instructions_sheet(wb["Instructions & Legend"], styles)

        logger.info("Populating Sheet 2: Overall A.8.10 Compliance...")
        populate_sheet2_overall(wb["Overall A.8.10 Compliance"], styles)

        logger.info("Populating Sheet 3: Retention Schedule Health...")
        populate_sheet3_retention(wb["Retention Schedule Health"], styles)

        logger.info("Populating Sheet 4: Deletion Method Effectiveness...")
        populate_sheet4_methods(wb["Deletion Method Effectiveness"], styles)

        logger.info("Populating Sheet 5: Third-Party Deletion Performance...")
        populate_sheet5_thirdparty(wb["Third-Party Deletion Performance"], styles)

        logger.info("Populating Sheet 6: Verification & Evidence Quality...")
        populate_sheet6_verification(wb["Verification & Evidence Quality"], styles)

        logger.info("Populating Sheet 7: Critical Gaps Dashboard...")
        populate_sheet7_gaps(wb["Critical Gaps Dashboard"], styles)

        logger.info("Populating Sheet 8: Trend Analysis...")
        populate_sheet8_trends(wb["Trend Analysis"], styles)

        logger.info("Populating Sheet 9: Executive Summary...")
        populate_sheet9_executive(wb["Executive Summary"], styles)

        # Save workbook
        filename = f"{FILENAME_PREFIX}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("=" * 70)
        logger.info("DASHBOARD GENERATION COMPLETE!")
        logger.info("=" * 70)
        logger.info("File: %s", filename)
        logger.info("Version: %s", VERSION)
        logger.info("Structure: 9 sheets (Instructions + 7 Dashboard sections + Executive Summary)")
        logger.info("  - External links to normalized source files")
        logger.info("  - Auto-populates metrics (no manual data entry for key metrics)")
        logger.info("Cell Reference Mapping (VERIFIED):")
        logger.info("  A.8.10.1-3: 'Summary Dashboard'!G10 (Compliance %%)")
        logger.info("  A.8.10.4: 'Verification Dashboard'!B9 (Compliance %%) - DIFFERENT SHEET NAME")
        logger.info("=" * 70)
        return 0
    except Exception as e:
        logger.error("Failed to generate dashboard: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
