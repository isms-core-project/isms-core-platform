#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.10.1 - Retention & Deletion Triggers Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.10: Information Deletion
Assessment Domain 1 of 4: Retention Schedule & Deletion Trigger Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific data categories, retention requirements, regulatory
obligations, and deletion trigger mechanisms.

Key customisation areas:
1. Data category taxonomy (match your organisation's data classification scheme)
2. Retention periods (adapt to your regulatory and business requirements)
3. Deletion trigger types (automated vs. manual trigger mechanisms)
4. Legal hold procedures (specific to your legal/compliance processes)
5. Data subject rights workflows (GDPR/FADP compliance procedures)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.10 Information Deletion Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
retention schedules and deletion trigger mechanisms against ISO 27001:2022 Control
A.8.10 requirements, supporting evidence-based validation of systematic information
deletion practices across the data lifecycle.

**Purpose:**
Enables systematic assessment of retention schedule definition, documentation,
and enforcement mechanisms to ensure information is deleted when no longer required
for business, legal, or regulatory purposes.

**Assessment Scope:**
- Complete data category inventory and classification
- Retention period definition and regulatory alignment
- Automated and manual deletion trigger configuration
- Legal hold management procedures and tracking
- Data subject rights (GDPR Article 17 / Swiss FADP erasure requests)
- Retention schedule documentation and approval
- Deletion trigger testing and verification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance, regulatory context, color coding
2. Data Category Registry - Complete inventory of all data types and classifications
3. Retention Sched. Compliance - Legal/regulatory alignment for retention periods
4. Deletion Trigger Config. - Automated and manual deletion mechanisms
5. Legal Hold Management - Procedures to suspend deletion when legally required
6. Data Subject Rights - GDPR/FADP erasure request handling workflows
7. Summary Dashboard - Compliance overview, KPIs, critical gaps identification
8. Evidence Register - Audit evidence tracking and documentation linkage
9. Approval Sign-Off - Three-level stakeholder approval workflow

**Key Features:**
- Data validation with data classification and retention period dropdown lists
- Conditional formatting for retention compliance status indication
- Automated gap identification for missing retention schedules
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with GDPR Article 17 and Swiss FADP requirements
- Legal hold tracking and suspension procedures

**Integration:**
consolidates data from all four information deletion assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a810_1_retention_triggers.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a810_1_retention_triggers.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a810_1_retention_triggers.py --date 20250124

Output:
    File: ISMS_A_8_10_1_Retention_Deletion_Triggers_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize retention periods to match your regulatory context
    2. Inventory all data categories across all systems and processes
    3. Complete retention schedule for each data category
    4. Validate deletion triggers are properly configured
    5. Review legal hold procedures and tracking mechanisms
    6. Document data subject rights handling workflows
    7. Conduct gap analysis for missing or incomplete retention schedules
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (retention policies, deletion logs)
    10. Obtain stakeholder approvals

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.10
Assessment Domain:    1 of 4 (Retention Schedule & Deletion Trigger Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.10: Information Deletion Policy (Governance)
    - ISMS-IMP-A.8.10.1: Retention & Deletion Triggers Implementation Guide
    - ISMS-IMP-A.8.10.2: Deletion Methods Assessment (Domain 2)
    - ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion Assessment (Domain 3)
    - ISMS-IMP-A.8.10.4: Verification & Evidence Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.10.1 specification
    - Supports comprehensive retention schedule and deletion trigger evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Regulatory Requirements:**
Retention periods must comply with applicable regulations including:
- GDPR: Article 5(1)(e) storage limitation principle
- Swiss FADP: Article 6(3) data processing principles (proportionality)
- Industry-specific: PCI DSS v4.0.1 (payment data), HIPAA (health records), etc.
- Jurisdiction-specific: Local data protection and archival laws

This assessment provides a framework - customize retention periods based on
your organisation's legal, regulatory, and business requirements.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of retention schedule documentation, deletion
trigger configuration, and legal hold procedures.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Data category inventory and classification
- Retention periods and business justifications
- System inventory and data locations
- Legal hold information

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new data categories or regulatory changes
- Semi-annually: Update retention periods for legislative changes
- Annually: Complete reassessment of all data categories and deletion triggers
- Ad-hoc: When new systems are deployed or regulatory requirements change

**Quality Assurance:**
Have Data Protection Officers, Legal/Compliance teams, and Business Unit Owners
validate assessments before using results for compliance reporting or remediation
decisions.

**Regulatory Alignment:**
Ensure retention periods align with applicable regulatory requirements:
- Privacy laws: GDPR Article 17, Swiss FADP erasure requirements
- Financial services: Basel III, MiFID II, Dodd-Frank retention requirements
- Healthcare: HIPAA minimum necessary and retention requirements
- Payment processing: PCI DSS v4.0.1 data retention and disposal requirements
- Public sector: Government archival and public records laws

Customize assessment criteria to include regulatory-specific requirements.

**Data Subject Rights (GDPR Article 17 / Swiss FADP):**
Organisations must be able to:
1. Identify all systems containing a data subject's information
2. Validate the identity of the requestor
3. Assess whether deletion exceptions apply (legal obligations, public interest)
4. Execute deletion across all systems within required timeframes
5. Document what was deleted and what was retained (with justification)
6. Notify third parties who received the data
7. Respond to the data subject with confirmation

This assessment evaluates the organisation's capability to fulfill these
obligations systematically and within regulatory timeframes.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.10.1"
WORKBOOK_NAME = "Retention & Deletion Triggers Assessment"
CONTROL_ID = "A.8.10"
CONTROL_NAME = "Information Deletion"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠ Warning sign
BULLET = '\u2022'     # * Bullet point
ARROW = '\u2192'      # -> Right arrow

# ==========================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ==========================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Create all sheets in order
    wb.create_sheet("Instructions & Legend", 0)
    wb.create_sheet("2. Data Category Registry", 1)
    wb.create_sheet("3. Retention Sched. Compliance", 2)
    wb.create_sheet("4. Deletion Trigger Config.", 3)
    wb.create_sheet("5. Legal Hold Management", 4)
    wb.create_sheet("6. Data Subject Requests", 5)
    wb.create_sheet("Evidence Register", 6)
    wb.create_sheet("Summary Dashboard", 7)
    wb.create_sheet("Approval Sign-Off", 8)
    
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    styles = {
        'title': {
            'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'header': {
            'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'subheader': {
            'font': Font(name='Calibri', size=10, bold=True, color='000000'),
            'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'input_cell': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_compliant': {
            'font': Font(name='Calibri', size=10, bold=True, color='006100'),
            'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_partial': {
            'font': Font(name='Calibri', size=10, bold=True, color='9C5700'),
            'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_noncompliant': {
            'font': Font(name='Calibri', size=10, bold=True, color='C00000'),
            'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'reference_cell': {
            'font': Font(name='Calibri', size=9, color='000000'),
            'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'normal': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    }
    return styles


# ==========================================================================
# SECTION 2: COLUMN DEFINITIONS
# ==========================================================================


_STYLES = setup_styles()
def get_base_columns():
    """Return base column structure (A-Q, 17 columns)."""
    return [
        ("A", "Data Category / System Name", 30),
        ("B", "Data Classification", 22),
        ("C", "Business Owner", 18),
        ("D", "Retention Period", 20),
        ("E", "Legal/Regulatory Basis", 20),
        ("F", "Status", 18),
        ("G", "Implementation Date", 12),
        ("H", "Last Review Date", 12),
        ("I", "Next Review Date", 12),
        ("J", "Gap Identified", 25),
        ("K", "Remediation Plan", 25),
        ("L", "Target Completion", 12),
        ("M", "Risk Level", 12),
        ("N", "Evidence Reference", 20),
        ("O", "Notes / Comments", 25),
        ("P", "Remediation Owner", 18),
        ("Q", "Budget Required", 15)
    ]


def get_extended_columns_sheet2():
    """Sheet 2: Data Category Registry extensions."""
    return [
        ("R", "Primary Storage Location", 22),
        ("S", "Volume/Records", 18),
        ("T", "Contains PII/SPI", 18)
    ]


def get_extended_columns_sheet3():
    """Sheet 3: Retention Schedule Compliance extensions."""
    return [
        ("R", "Retention Calculation Method", 22),
        ("S", "Event Trigger Description", 25),
        ("T", "Backup Retention Aligned", 18)
    ]


def get_extended_columns_sheet4():
    """Sheet 4: Deletion Trigger Configuration extensions."""
    return [
        ("R", "Trigger Type", 18),
        ("S", "Trigger Frequency", 18),
        ("T", "Legal Hold Check Integrated", 22)
    ]


def get_extended_columns_sheet5():
    """Sheet 5: Legal Hold Management extensions."""
    return [
        ("R", "Active Legal Holds", 15),
        ("S", "Legal Hold Notification Process", 25),
        ("T", "Hold Review Frequency", 18)
    ]


def get_extended_columns_sheet6():
    """Sheet 6: Data Subject Requests extensions."""
    return [
        ("R", "Average Response Time (Days)", 20),
        ("S", "GDPR/FADP Applicable", 18),
        ("T", "Request Volume (Last 12 Months)", 22)
    ]


# ==========================================================================
# SECTION 3: DATA VALIDATION
# ==========================================================================

def create_base_validations(ws):
    """Create data validation objects for standard columns."""
    validations = []

    # Data Classification (Column B)
    dv_classification = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=True
    )
    dv_classification.error = 'Please select from dropdown'
    dv_classification.errorTitle = 'Invalid Classification'
    dv_classification.add('B10:B100')
    validations.append(dv_classification)

    # Retention Period (Column D)
    dv_retention = DataValidation(
        type="list",
        formula1='"30 days,60 days,90 days,6 months,1 year,2 years,3 years,5 years,7 years,10 years,Permanent,Until Event Occurs,Other (specify in notes)"',
        allow_blank=True
    )
    dv_retention.error = 'Please select from dropdown'
    dv_retention.errorTitle = 'Invalid Retention Period'
    dv_retention.add('D10:D100')
    validations.append(dv_retention)

    # Legal/Regulatory Basis (Column E)
    dv_legal = DataValidation(
        type="list",
        formula1='"Swiss FADP,EU GDPR,Swiss Code of Obligations (OR),Swiss Tax Law,EU ePrivacy Directive,Industry Standard (specify),Contractual Obligation,Legitimate Interest,Consent,Legal Obligation,Multiple Bases (specify),Other (specify in notes)"',
        allow_blank=True
    )
    dv_legal.error = 'Please select from dropdown'
    dv_legal.errorTitle = 'Invalid Legal Basis'
    dv_legal.add('E10:E100')
    validations.append(dv_legal)

    # Status (Column F)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=True
    )
    dv_status.error = 'Please select from dropdown'
    dv_status.errorTitle = 'Invalid Status'
    dv_status.add('F10:F100')
    validations.append(dv_status)

    # Risk Level (Column M)
    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    dv_risk.error = 'Please select from dropdown'
    dv_risk.errorTitle = 'Invalid Risk Level'
    dv_risk.add('M10:M100')
    validations.append(dv_risk)

    # Budget Required (Column Q)
    dv_budget = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    dv_budget.error = 'Please select from dropdown'
    dv_budget.errorTitle = 'Invalid Budget Option'
    dv_budget.add('Q10:Q100')
    validations.append(dv_budget)

    for _dv in validations:
        ws.add_data_validation(_dv)

def create_sheet_specific_validations(ws, sheet_type):
    """Create sheet-specific data validations."""
    validations = []

    if sheet_type == "sheet2":
        # Primary Storage Location (Column R)
        dv = DataValidation(
            type="list",
            formula1='"On-Premise,Cloud (IaaS),Cloud (PaaS),Cloud (SaaS),Hybrid,Third-Party"',
            allow_blank=True
        )
        dv.add('R10:R100')
        validations.append(dv)

        # Contains PII/SPI (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes - PII,Yes - SPI,Yes - Both,No"',
            allow_blank=True
        )
        dv.add('T10:T100')
        validations.append(dv)

    elif sheet_type == "sheet3":
        # Retention Calculation Method (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Fixed Period,Event-Based,Hybrid"',
            allow_blank=True
        )
        dv.add('R10:R100')
        validations.append(dv)

        # Backup Retention Aligned (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes,No,Partial,N/A"',
            allow_blank=True
        )
        dv.add('T10:T100')
        validations.append(dv)

    elif sheet_type == "sheet4":
        # Trigger Type (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Automatic,Manual,Semi-Automatic,Event-Based"',
            allow_blank=True
        )
        dv.add('R10:R100')
        validations.append(dv)

        # Trigger Frequency (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Real-time,Daily,Weekly,Monthly,Quarterly,Annual,Ad-hoc"',
            allow_blank=True
        )
        dv.add('S10:S100')
        validations.append(dv)

        # Legal Hold Check Integrated (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Automated,Yes - Manual,No,N/A"',
            allow_blank=True
        )
        dv.add('T10:T100')
        validations.append(dv)

    elif sheet_type == "sheet5":
        # Legal Hold Notification Process (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Automated,Manual,Hybrid,None"',
            allow_blank=True
        )
        dv.add('S10:S100')
        validations.append(dv)

        # Hold Review Frequency (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Weekly,Monthly,Quarterly,Annual"',
            allow_blank=True
        )
        dv.add('T10:T100')
        validations.append(dv)

    elif sheet_type == "sheet6":
        # GDPR/FADP Applicable (Column S)
        dv = DataValidation(
            type="list",
            formula1='"GDPR Only,FADP Only,Both,Neither"',
            allow_blank=True
        )
        dv.add('S10:S100')
        validations.append(dv)

    for _dv in validations:
        ws.add_data_validation(_dv)

# ==========================================================================
# SECTION 4: HELPER FUNCTIONS
# ==========================================================================

def apply_cell_style(cell, styles, style_type):
    """Apply style dictionary to a cell."""
    style = styles[style_type]
    cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


def create_header_row(ws, row, columns, styles):
    """Create header row with column definitions."""
    for col_letter, header_text, width in columns:
        cell = ws[f'{col_letter}{row}']
        cell.value = header_text
        apply_cell_style(cell, styles, 'header')
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[row].height = 35


def create_data_rows(ws, start_row, num_rows, num_cols, styles):
    """Create yellow-highlighted data entry rows."""
    for row in range(start_row, start_row + num_rows):
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles, 'input_cell')


def create_checklist_section(ws, start_row, checklist_items, styles):
    """Create compliance checklist section."""
    # Section header
    ws.merge_cells(f'A{start_row}:T{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = f"COMPLIANCE CHECKLIST - Mark ✅ or {XMARK}"
    apply_cell_style(cell, styles, 'subheader')
    
    # Checklist items
    current_row = start_row + 1
    for idx, item in enumerate(checklist_items, 1):
        # Checkbox column
        ws[f'A{current_row}'].value = f"{idx}."
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
        ws.column_dimensions['A'].width = 5
        
        # Item description
        ws.merge_cells(f'B{current_row}:R{current_row}')
        ws[f'B{current_row}'].value = item
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Status column
        ws[f'S{current_row}'].value = ""
        ws[f'S{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'S{current_row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.column_dimensions['S'].width = 8
        
        # Notes column
        ws[f'T{current_row}'].value = ""
        ws[f'T{current_row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        current_row += 1
    
    return current_row


def create_reference_table(ws, start_row, table_title, headers, data, styles):
    """Create reference table with title and data."""
    # Table title
    num_cols = len(headers)
    end_col = get_column_letter(num_cols)
    ws.merge_cells(f'A{start_row}:{end_col}{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = table_title
    apply_cell_style(cell, styles, 'subheader')
    
    # Headers
    header_row = start_row + 1
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{header_row}']
        cell.value = header
        apply_cell_style(cell, styles, 'header')
    
    # Data rows
    current_row = header_row + 1
    for row_data in data:
        for col_idx, value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = value
            apply_cell_style(cell, styles, 'reference_cell')
        current_row += 1
    
    return current_row + 1  # Return next available row


# ==========================================================================
# SECTION 5: ASSESSMENT SHEET CREATOR (CORE FUNCTION)
# ==========================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, 
                            assessment_question, base_cols, extended_cols,
                            checklist_items, reference_tables, sheet_type):
    """
    Create standardized assessment sheet with:
    - Title and policy reference
    - Assessment question
    - Column headers
    - Data entry rows (13 rows, yellow-highlighted)
    - Compliance checklist
    - Reference tables
    - Exception/deviation block
    """
    
    # Row 1: Title
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = section_title.upper()
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 35

    # Row 2: Policy Reference
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = f"Policy Reference: {policy_ref}"
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='left')

    # Row 4-6: Assessment Question
    ws.merge_cells('A4:T6')
    cell = ws['A4']
    cell.value = f"ASSESSMENT QUESTION:\n{assessment_question}"
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Row 7: Instructions
    ws.merge_cells('A7:T7')
    cell = ws['A7']
    cell.value = "Complete the yellow-highlighted cells below. Use dropdowns where provided. Link evidence in Column N to Evidence Register sheet."
    cell.font = Font(name='Calibri', size=9, italic=True, color='4472C4')
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    # Row 9: Column Headers (Base + Extended)
    all_columns = base_cols + extended_cols
    create_header_row(ws, 9, all_columns, styles)
    
    # Rows 10-22: Data Entry Rows (13 rows)
    create_data_rows(ws, 10, 13, len(all_columns), styles)
    
    # Rows 25+: Compliance Checklist
    next_row = create_checklist_section(ws, 25, checklist_items, styles)

    next_row += 1
    
    # Reference Tables
    for table_title, headers, data in reference_tables:
        next_row = create_reference_table(ws, next_row, table_title, headers, data, styles)
    
    # Exception/Deviation Block
    ws.merge_cells(f'A{next_row}:T{next_row}')
    cell = ws[f'A{next_row}']
    cell.value = "EXCEPTIONS / DEVIATIONS"
    apply_cell_style(cell, styles, 'subheader')
    
    next_row += 1
    ws.merge_cells(f'A{next_row}:T{next_row+2}')
    cell = ws[f'A{next_row}']
    cell.value = "Document any exceptions or deviations from requirements here (e.g., data categories excluded from assessment, pending reviews, disputed retention periods):"
    cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Freeze panes at row 10 (headers stay visible)
    ws.freeze_panes = 'A10'
    
    # Apply validations
    create_base_validations(ws)
    create_sheet_specific_validations(ws, sheet_type)


# ==========================================================================
# SECTION 6: INSTRUCTIONS SHEET
# ==========================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete assessment sheets 2–6 in order, filling in all yellow-highlighted cells.', '2. Use dropdown menus where provided for consistency.', '3. Link supporting evidence to the Evidence Register using Column N.', '4. Mark checklist items as compliant or non-compliant.', '5. Review the Summary Dashboard for overall compliance status.', '6. Document any gaps with remediation plans and target completion dates.', '7. Obtain approval using the Approval Sign-Off sheet.', '8. Review and update this assessment annually or when significant changes occur.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet -- standard common sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 -- merge A1:H1, title "EVIDENCE REGISTER"
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2 -- subtitle, italic
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers Row 4 -- 8 columns, 003366 fill, white bold font
    headers = [
        "Evidence ID", "Category", "Description", "Source Document",
        "Date Collected", "Collected By", "Status", "Notes"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, name="Calibri", color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Status dropdown (column G)
    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True
    )

    # Category dropdown (column B)
    cat_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Procedure Document,Screenshot,System Log Export,Configuration File,Email Communication,Meeting Minutes,Audit Report,Certificate,Contract/Agreement,Training Record,Test Result,Other"',
        allow_blank=True
    )

    validations = [status_dv, cat_dv]

    # Data rows 5-104 (100 rows) with EV-001 to EV-100
    # Apply FFFFCC to ALL columns (1-8) including Evidence ID column
    for r in range(5, 105):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            if c == 1:
                cell.value = f"EV-{r - 4:03d}"
                cell.font = Font(color="808080", name="Calibri")
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Apply dropdowns to ranges
    cat_dv.add("B5:B104")
    status_dv.add("G5:G104")

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 30

    ws.freeze_panes = "A5"


# ==========================================================================
# SECTION 8: APPROVAL SIGN-OFF
# ==========================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet -- standard common sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header Row 1 -- merge A1:E1
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2 -- subtitle banner
    ws.merge_cells("A2:E2")
    ws["A2"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A2"].font = Font(italic=True, name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border
    # Row 3: ASSESSMENT SUMMARY banner -- 4472C4 fill
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
    ]

    row += 1
    status_row = None
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if label == "Assessment Status:":
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Status dropdown on Assessment Status
    validations = []
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True
    )
    if status_row:
        status_dv.add(ws[f"B{status_row}"])
    validations.append(status_dv)

    # 3 approver sections
    approvers = [
        ("COMPLETED BY (DATA PROTECTION)", "4472C4"),
        ("REVIEWED BY (COMPLIANCE)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True, name="Calibri")
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL ASSESSMENT DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    decision_dv.add(ws[f"B{row}"])
    validations.append(decision_dv)

    # NEXT REVIEW DETAILS -- gap of 3 rows
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    for _dv in validations:
        ws.add_data_validation(_dv)

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    ws.freeze_panes = "A3"


# ==========================================================================
# SECTION 9: SUMMARY DASHBOARD
# ==========================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 format."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ffffcc = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    d9d9d9 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # ── A1:G1 title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    title.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── A2:G2 subtitle ───────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    subtitle = ws["A2"]
    subtitle.value = f"Summary Dashboard  |  {WORKBOOK_NAME}  |  Generated: {GENERATED_TIMESTAMP}"
    subtitle.font = Font(name="Calibri", size=11, italic=True, color="003366")
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"

    # =========================================================================
    # TABLE 1 — COMPLIANCE ASSESSMENT SUMMARY (rows 4 onward)
    # =========================================================================
    # Banner row 4
    ws.merge_cells("A4:G4")
    banner = ws["A4"]
    banner.value = "TABLE 1 \u2013 COMPLIANCE ASSESSMENT SUMMARY"
    banner.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    banner.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    banner.alignment = Alignment(horizontal="left", vertical="center")

    # Header row 5
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
                  "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows 6-10 (5 assessment areas) — col F is Status on each sheet
    # Sheets use col F DV: "✅ Compliant", "⚠️ Partial", "❌ Non-Compliant", "N/A"
    # Data rows start at row 10 on each sheet, run to row 100
    sheet_map = [
        ("Data Category Registry",      "'2. Data Category Registry'!F10:F100"),
        ("Retention Schedule Compliance","'3. Retention Sched. Compliance'!F10:F100"),
        ("Deletion Trigger Config.",     "'4. Deletion Trigger Config.'!F10:F100"),
        ("Legal Hold Management",        "'5. Legal Hold Management'!F10:F100"),
        ("Data Subject Requests",        "'6. Data Subject Requests'!F10:F100"),
    ]

    CHECK_VAL = "\u2705 Compliant"
    PARTIAL_VAL = "\u26a0\ufe0f Partial"
    NC_VAL = "\u274c Non-Compliant"
    NA_VAL = "N/A"

    row = 6
    for area_name, rng in sheet_map:
        c_formula  = f'=COUNTIF({rng},"{CHECK_VAL}")'
        p_formula  = f'=COUNTIF({rng},"{PARTIAL_VAL}")'
        nc_formula = f'=COUNTIF({rng},"{NC_VAL}")'
        na_formula = f'=COUNTIF({rng},"{NA_VAL}")'

        cells_in_row = [
            (ws.cell(row=row, column=1), area_name),
            (ws.cell(row=row, column=2), f"=C{row}+D{row}+E{row}+F{row}"),
            (ws.cell(row=row, column=3), c_formula),
            (ws.cell(row=row, column=4), p_formula),
            (ws.cell(row=row, column=5), nc_formula),
            (ws.cell(row=row, column=6), na_formula),
            (ws.cell(row=row, column=7), f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"),
        ]
        for cell, val in cells_in_row:
            cell.value = val
            cell.font = Font(name="Calibri", size=10, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=7).number_format = "0.0%"
        row += 1

    # TOTAL row
    total_row = row
    total_data = [
        (ws.cell(row=total_row, column=1), "TOTAL"),
        (ws.cell(row=total_row, column=2), f"=SUM(B6:B{total_row - 1})"),
        (ws.cell(row=total_row, column=3), f"=SUM(C6:C{total_row - 1})"),
        (ws.cell(row=total_row, column=4), f"=SUM(D6:D{total_row - 1})"),
        (ws.cell(row=total_row, column=5), f"=SUM(E6:E{total_row - 1})"),
        (ws.cell(row=total_row, column=6), f"=SUM(F6:F{total_row - 1})"),
        (ws.cell(row=total_row, column=7),
         f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"),
    ]
    for cell, val in total_data:
        cell.value = val
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=7).number_format = "0.0%"
    row = total_row + 2  # blank spacer

    # =========================================================================
    # TABLE 2 — KEY METRICS
    # =========================================================================
    ws.merge_cells(f"A{row}:G{row}")
    t2_banner = ws[f"A{row}"]
    t2_banner.value = "TABLE 2 \u2013 KEY METRICS"
    t2_banner.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    t2_banner.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    t2_banner.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    for col_idx, hdr in enumerate(["Metric", "Value", "Target"], start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    row += 1

    t2_data = [
        ("Overall Compliance Rate",
         f"=IFERROR(G{total_row},\"N/A\")",
         "\u2265 80%"),
        ("Data Categories with Defined Retention Schedule",
         "=IFERROR(COUNTIF('3. Retention Sched. Compliance'!F10:F100,\""
         + CHECK_VAL + "\"),0)",
         "100%"),
        ("Deletion Triggers Configured",
         "=IFERROR(COUNTIF('4. Deletion Trigger Config.'!F10:F100,\""
         + CHECK_VAL + "\"),0)",
         "100%"),
        ("Legal Hold Management Controls Compliant",
         "=IFERROR(COUNTIF('5. Legal Hold Management'!F10:F100,\""
         + CHECK_VAL + "\"),0)",
         "100%"),
        ("Data Subject Request Workflows Compliant",
         "=IFERROR(COUNTIF('6. Data Subject Requests'!F10:F100,\""
         + CHECK_VAL + "\"),0)",
         "100%"),
        ("Non-Compliant Items Requiring Remediation",
         f"=IFERROR(SUM(E6:E{total_row - 1}),0)",
         "0"),
    ]

    for metric, value, target in t2_data:
        ws.cell(row=row, column=1).value = metric
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=3).value = target
        for col_idx in range(1, 4):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = ffffcc
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).number_format = "0.0%"
        row += 1

    row += 1  # spacer

    # =========================================================================
    # TABLE 3 — KEY FINDINGS & RECOMMENDATIONS
    # =========================================================================
    ws.merge_cells(f"A{row}:G{row}")
    t3_banner = ws[f"A{row}"]
    t3_banner.value = "TABLE 3 \u2013 KEY FINDINGS & RECOMMENDATIONS"
    t3_banner.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    t3_banner.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    t3_banner.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    t3_headers = ["#", "Finding", "Impact", "Recommendation", "Priority", "Status", "Notes"]
    for col_idx, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = d9d9d9
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    row += 1

    t3_rows = [
        ("1", "Retention schedules may not cover all data categories",
         "High", "Complete data category inventory and assign retention periods to all items",
         "P1", "Open", ""),
        ("2", "Deletion trigger automation may be incomplete",
         "High", "Implement automated deletion triggers for all high-volume data categories",
         "P1", "Open", ""),
        ("3", "Legal hold integration with deletion systems requires verification",
         "Medium", "Test legal hold suspension across all automated deletion trigger systems",
         "P2", "Open", ""),
        ("4", "Data subject erasure request response times need monitoring",
         "Medium", "Establish SLA tracking and escalation for GDPR/FADP erasure requests",
         "P2", "Open", ""),
    ]

    for row_data in t3_rows:
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.fill = ffffcc
            cell.border = border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
        row += 1


# ==========================================================================
# SECTION 10: DOMAIN-SPECIFIC SHEET CREATORS
# ==========================================================================

def create_sheet2_data_category_registry(ws, styles):
    """Create Sheet 2: Data Category Registry."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet2()
    
    checklist = [
        "All business units have been consulted for data categories",
        "Data inventory includes operational, HR, financial, and customer data",
        "Each data category has a designated Business Owner",
        "Storage locations (on-premise, cloud, third-party) are documented",
        "PII and SPI data categories are specifically identified",
        "Volume/scale of data is estimated for each category",
        "Data classification levels are assigned per organisational policy",
        "Temporary/ephemeral data categories are included",
        "Legacy systems and archived data are included in inventory",
        "Cross-border data transfers are noted (if applicable)",
        "Data category registry is reviewed at least annually",
        "New business processes trigger data category reassessment",
        "Shadow IT data categories have been investigated",
        "M&A due diligence includes data category mapping",
        "Data categories align with Records Retention Schedule"
    ]
    
    reference_tables = [
        (
            "Common Data Categories",
            ["Category Example", "Typical Retention", "Common Legal Basis"],
            [
                ["Employee Personnel Files", "10 years post-employment", "Swiss OR, Tax Law"],
                ["Customer Contracts", "10 years post-termination", "Swiss OR (Art. 127)"],
                ["Financial Records", "10 years", "Swiss Tax Law"],
                ["Marketing Consent Records", "Until consent withdrawn + 1 year", "GDPR, FADP"],
                ["Access Logs", "1 year", "Security requirement"],
                ["Email Communications", "1-7 years (varies)", "Business need"],
                ["Customer Support Tickets", "3 years post-resolution", "Contractual"],
                ["Product Telemetry", "90 days - 2 years", "Legitimate interest"],
                ["CCTV Footage", "30-90 days", "Security/legal requirement"],
                ["Website Analytics", "26 months (GDPR default)", "Consent/Legitimate interest"],
                ["Backup Data", "Aligned with active retention", "Same as source data"],
                ["Development/Test Data", "Until project end + 90 days", "Business need"],
                ["Audit Trails", "7-10 years", "Regulatory requirement"],
                ["HR Recruitment Records", "6 months post-decision", "GDPR, FADP"],
                ["Health & Safety Records", "10 years", "Legal obligation"]
            ]
        ),
        (
            "Storage Location Implications",
            ["Location Type", "Deletion Complexity", "Key Considerations"],
            [
                ["On-Premise", "Medium", "Full control, hardware lifecycle"],
                ["Cloud (IaaS)", "Medium-High", "Snapshot management, API deletion"],
                ["Cloud (PaaS)", "High", "Platform-specific deletion methods"],
                ["Cloud (SaaS)", "Very High", "Vendor-dependent, see A.8.10.3"],
                ["Hybrid", "Very High", "Coordination across environments"],
                ["Third-Party", "Very High", "Contractual deletion, see A.8.10.3"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 2: Data Category Registry",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.1",
        "Have we identified and documented ALL data categories we process, including their characteristics, ownership, and storage locations?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet2"
    )


def create_sheet3_retention_schedule(ws, styles):
    """Create Sheet 3: Retention Schedule Compliance."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet3()
    
    checklist = [
        "Every data category has a defined retention period",
        "Retention periods have documented legal/regulatory basis",
        "Swiss FADP compliance verified for all PII categories",
        "GDPR compliance verified where applicable",
        "Retention periods do not exceed legal maximums",
        "Retention periods meet legal minimums (e.g., tax records)",
        "Event-based retention triggers are clearly defined",
        "Business justification documented for non-regulatory retention",
        "Conflicting retention requirements are resolved and documented",
        "Retention schedule approved by Legal/Compliance team",
        "Backup retention aligns with active data retention",
        "'Permanent' retention is justified and approved",
        "Retention periods reviewed annually",
        "Changes to retention periods follow change management process",
        "Data subjects are informed of retention periods (GDPR/FADP)"
    ]
    
    reference_tables = [
        (
            "Swiss Legal Retention Requirements",
            ["Data Type", "Minimum Retention", "Legal Basis", "Maximum Deletion"],
            [
                ["Accounting Records", "10 years", "Swiss OR Art. 958f", "After 10 years"],
                ["Tax Records", "10 years", "Swiss Tax Law", "After 10 years"],
                ["Employee Records", "10 years", "Swiss OR Art. 127", "After 10 years (FADP)"],
                ["Contracts", "10 years", "Swiss OR Art. 127", "After 10 years"],
                ["Payroll Records", "10 years", "Swiss Tax/Social Law", "After 10 years"],
                ["Corporate Minutes", "Permanent", "Swiss OR Art. 695", "N/A"],
                ["Share Register", "Permanent", "Swiss OR Art. 686", "N/A"],
                ["Correspondence", "No minimum", "Business need", "FADP: when no longer needed"],
                ["Marketing Data", "No minimum", "Consent-based", "FADP: consent withdrawal"],
                ["CCTV Footage", "No minimum", "Proportionality", "FADP: typically 30-90 days"]
            ]
        ),
        (
            "Retention Calculation Methods",
            ["Method", "Definition", "Example"],
            [
                ["Fixed Period", "Set duration from creation/collection", "7 years from date of invoice"],
                ["Event-Based", "Duration starts at specific event", "10 years from contract termination"],
                ["Hybrid", "Combination of fixed + event", "3 years OR until litigation resolved"],
                ["Indefinite (Permanent)", "No planned deletion", "Corporate charter documents"],
                ["Until Consent Withdrawn", "Active until user action", "Marketing email list"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 3: Retention Schedule Compliance",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.2",
        "Does each data category have a documented retention period with a valid legal or business justification?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet3"
    )


def create_sheet4_deletion_triggers(ws, styles):
    """Create Sheet 4: Deletion Trigger Configuration."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet4()
    
    checklist = [
        "Automated deletion triggers implemented for high-volume data",
        "Manual deletion procedures documented for low-volume data",
        "Trigger execution is logged and auditable",
        "Failed deletion attempts are alerted and investigated",
        "Legal hold checks are integrated into trigger logic",
        "Backup deletion triggers aligned with active data triggers",
        "Cloud/SaaS deletion triggers verified (see A.8.10.3)",
        "Event-based triggers are monitored for event occurrence",
        "Deletion triggers tested at least annually",
        "Trigger failures have defined escalation procedures",
        "Business Owner approval required for manual triggers (if policy requires)",
        "Deletion triggers respect data dependencies (referential integrity)",
        "Triggers account for data in transit (queues, caches)",
        "Deletion trigger schedule published and communicated",
        "Trigger configuration changes follow change management"
    ]
    
    reference_tables = [
        (
            "Trigger Types and Use Cases",
            ["Trigger Type", "Best For", "Advantages", "Disadvantages"],
            [
                ["Automatic", "High-volume, structured data", "Consistent, no human error", "Requires robust testing"],
                ["Manual", "Low-volume, sensitive data", "Controlled, deliberate", "Human error risk, slower"],
                ["Semi-Automatic", "Medium volume, needs approval", "Balance control and efficiency", "Workflow complexity"],
                ["Event-Based", "Contract-dependent data", "Precise, legally aligned", "Requires event tracking"]
            ]
        ),
        (
            "Common Deletion Trigger Patterns",
            ["Pattern", "Description", "Example Implementation"],
            [
                ["Time-Based Cron", "Scheduled deletion job", "Daily at 02:00 UTC, delete records > 7 years"],
                ["Database TTL", "Built-in expiration", "MongoDB TTL index, Redis EXPIRE"],
                ["Lifecycle Policy", "Cloud-native deletion", "AWS S3 Lifecycle, Azure Blob Lifecycle"],
                ["Workflow Approval", "Human-in-the-loop", "Jira ticket → approval → script execution"],
                ["Event Listener", "Reactive deletion", "Contract.status = terminated → 90-day timer → delete"],
                ["Batch Processing", "Periodic bulk deletion", "Monthly job: identify + delete expired records"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 4: Deletion Trigger Configuration",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.3",
        "Do we have effective and reliable deletion triggers that ensure data is deleted when retention periods expire?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet4"
    )


def create_sheet5_legal_hold(ws, styles):
    """Create Sheet 5: Legal Hold Management."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet5()
    
    checklist = [
        "Legal hold procedures are documented and approved",
        "Legal/Compliance team has authority to initiate holds",
        "IT/Data teams receive timely notification of holds",
        "Hold notification includes specific data categories and date ranges",
        "Automated deletion triggers check for active holds",
        "Legal hold data is segregated or flagged in systems",
        "Legal hold registry is maintained (who, what, when, why)",
        "Business Owners are notified when their data is under hold",
        "Periodic review of active holds (at least quarterly)",
        "Hold release procedures are documented",
        "Post-hold deletion is executed per normal retention schedule",
        "Legal hold training provided to relevant staf",
        "Legal hold breaches (accidental deletion) are reported and investigated",
        "Legal hold scope is clearly defined (not overly broad)",
        "Legal hold process tested at least annually"
    ]
    
    reference_tables = [
        (
            "Legal Hold Triggers",
            ["Trigger Event", "Typical Response Time", "Hold Scope"],
            [
                ["Litigation Notice", "Within 24 hours", "All relevant data for case"],
                ["Regulatory Investigation", "Within 48 hours", "Specified data types/periods"],
                ["Internal Investigation", "Within 72 hours", "Specific employees/systems"],
                ["M&A Due Diligence", "Within 1 week", "All corporate records"],
                ["Employment Dispute", "Within 48 hours", "Employee records + communications"],
                ["IP Dispute", "Within 72 hours", "Development records, communications"],
                ["Data Breach Investigation", "Immediate", "Affected systems/logs"],
                ["Audit Notification", "Within 1 week", "Records relevant to audit scope"]
            ]
        ),
        (
            "Hold Management Best Practices",
            ["Practice", "Purpose", "Implementation"],
            [
                ["Centralised Hold Registry", "Single source of truth", "Legal hold database/system"],
                ["Automated Flagging", "Prevent accidental deletion", "System flags or separate storage"],
                ["Regular Hold Reviews", "Release unnecessary holds", "Quarterly review meetings"],
                ["Clear Release Process", "Controlled return to normal deletion", "Legal approval required"],
                ["Documentation Requirements", "Audit trail", "Who, what, when, why, how long"],
                ["Scope Definition", "Avoid over-preservation", "Specific data types, date ranges"],
                ["Cost Tracking", "Understand hold impact", "Storage costs, resource time"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 5: Legal Hold Management",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.4",
        "Do we have robust processes to identify, communicate, and manage legal holds to prevent premature deletion of data under hold?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet5"
    )


def create_sheet6_data_subject_requests(ws, styles):
    """Create Sheet 6: Data Subject Requests (GDPR/FADP)."""
    
    base_cols = get_base_columns()
    extended_cols = get_extended_columns_sheet6()
    
    checklist = [
        "Data subject request procedure is documented",
        "Request intake channel is clearly communicated (e.g., privacy@company.ch)",
        "Identity verification process prevents fraudulent requests",
        "Request acknowledgment within 5 business days (GDPR/FADP)",
        "Requests fulfilled within 30 days (extendable to 60 days with justification)",
        "Search capability across all systems holding personal data",
        "Ability to distinguish between deletion obligations and exceptions",
        "Third-party/cloud systems included in deletion scope (see A.8.10.3)",
        "Backup deletion addressed in response (or timeline communicated)",
        "Request tracking system in place (ticket system, CRM, etc.)",
        "Responses documented (what was deleted, what was retained, why)",
        "Data subjects informed of retention exceptions (e.g., legal obligations)",
        "Escalation procedure for complex requests",
        "Staff training on data subject request handling",
        "Request metrics tracked and reported to DPO/Privacy Officer"
    ]
    
    reference_tables = [
        (
            "GDPR Article 17 Deletion Obligations",
            ["Condition for Deletion", "Organisation Must Delete If...", "Example"],
            [
                ["No longer necessary", "Data no longer needed for original purpose", "Marketing data after consent withdrawn"],
                ["Consent withdrawn", "Individual withdraws consent (no other legal basis)", "Newsletter subscription cancellation"],
                ["Objection to processing", "Individual objects + no overriding legitimate grounds", "Profiling/direct marketing objection"],
                ["Unlawful processing", "Data was processed illegally", "Processing without valid legal basis"],
                ["Legal obligation", "Law requires deletion", "Court order, regulatory requirement"],
                ["Child's data", "Data collected from child w/o proper consent", "Social media account created under age 13"]
            ]
        ),
        (
            "GDPR Article 17 Deletion Exceptions",
            ["Exception", "Organisation Can Retain If...", "Example"],
            [
                ["Freedom of expression", "Exercise of freedom of expression/information", "Journalistic/academic/artistic purposes"],
                ["Legal obligation", "Processing necessary for legal compliance", "Tax records (10-year retention)"],
                ["Public interest", "Public health, scientific research (safeguarded)", "Anonymized medical research data"],
                ["Legal claims", "Establishment, exercise, defence of legal claims", "Litigation hold, contract disputes"],
                ["Archiving in public interest", "Historical, statistical, scientific research (safeguarded)", "National archives, public health studies"]
            ]
        ),
        (
            "Request Response Workflow",
            ["Step", "Action", "Timeframe", "Responsible Role"],
            [
                ["1. Receive", "Log request in tracking system", "Day 0", "Privacy Team"],
                ["2. Acknowledge", "Send acknowledgment to data subject", "Within 5 business days", "Privacy Team"],
                ["3. Verify Identity", "Validate requestor identity", "Within 7 days", "Privacy Team"],
                ["4. Assess Scope", "Identify all systems with data subject's data", "Within 10 days", "IT + Business Owners"],
                ["5. Check Exceptions", "Determine if deletion exceptions apply", "Within 15 days", "Legal/Compliance"],
                ["6. Execute Deletion", "Delete data across all systems", "Within 25 days", "IT + Third-Parties"],
                ["7. Document", "Record what was deleted and retained", "Day 28", "Privacy Team"],
                ["8. Respond", "Inform data subject of outcome", "Day 30", "Privacy Team"],
                ["9. Follow-Up", "Backup deletion (if applicable)", "Within 90 days", "IT"]
            ]
        )
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 6: Data Subject Requests (GDPR/FADP)",
        "ISMS-POL-A.8.10-S2.1, Section 2.1.5",
        "Can we effectively identify, validate, and execute deletion requests from data subjects within required timeframes?",
        base_cols, extended_cols,
        checklist, reference_tables,
        "sheet6"
    )


# ==========================================================================
# SECTION 11: REFERENCE DATA
# ==========================================================================

# (Checklists and reference tables are defined inline in sheet creators above)


# ==========================================================================
# SECTION 12: MAIN EXECUTION
# ==========================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    try:
        logger.info("=" * 78)
        logger.info("ISMS-IMP-A.8.10.1 - Retention & Deletion Triggers Assessment Generator")
        logger.info("ISO/IEC 27001:2022 Control A.8.10: Information Deletion")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/9] Creating Instructions & Legend...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/9] Creating Sheet 2: Data Category Registry...")
        create_sheet2_data_category_registry(wb["2. Data Category Registry"], styles)

        logger.info("[3/9] Creating Sheet 3: Retention Schedule Compliance...")
        create_sheet3_retention_schedule(wb["3. Retention Sched. Compliance"], styles)

        logger.info("[4/9] Creating Sheet 4: Deletion Trigger Configuration...")
        create_sheet4_deletion_triggers(wb["4. Deletion Trigger Config."], styles)

        logger.info("[5/9] Creating Sheet 5: Legal Hold Management...")
        create_sheet5_legal_hold(wb["5. Legal Hold Management"], styles)

        logger.info("[6/9] Creating Sheet 6: Data Subject Requests...")
        create_sheet6_data_subject_requests(wb["6. Data Subject Requests"], styles)

        logger.info("[8/9] Creating Evidence Register (100 rows)...")
        create_evidence_register(wb["Evidence Register"], styles)

        logger.info("[7/9] Creating Summary Dashboard...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

        logger.info("[9/9] Creating Approval Sign-Off (3-level workflow)...")
        create_approval_sheet(wb["Approval Sign-Off"], styles)

        filename = f"ISMS-IMP-A.8.10.1_Retention_Deletion_Triggers_{datetime.now().strftime('%Y%m%d')}.xlsx"
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info("=" * 78)
        logger.info(f"SUCCESS: {filename}")
        logger.info("Workbook Structure:")
        logger.info("  - Instructions & Legend - Complete usage guide")
        logger.info("  - Sheet 2: Data Category Registry - Complete data inventory")
        logger.info("  - Sheet 3: Retention Schedule Compliance - Legal/regulatory alignment")
        logger.info("  - Sheet 4: Deletion Trigger Configuration - Automated and manual triggers")
        logger.info("  - Sheet 5: Legal Hold Management - Hold procedures and active holds")
        logger.info("  - Sheet 6: Data Subject Requests - GDPR Article 17 / FADP compliance")
        logger.info("  - Summary Dashboard - Executive overview with KPIs")
        logger.info("  - Evidence Register - 100 rows for supporting documentation")
        logger.info("  - Approval Sign-Off - 3-level approval workflow")
        logger.info("Key Features: Vendor-neutral, 13 data entry rows, compliance checklists")
        logger.info("Next Steps: Open workbook, complete yellow fields, review dashboard")
        logger.info("Related: A.8.10.2 (Deletion Methods), A.8.10.3 (Third-Party)")
        logger.info("=" * 78)
        return 0
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
