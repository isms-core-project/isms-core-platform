#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.23.2 - Network Coverage Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.23: Web Filtering
Assessment Domain 2 of 4: Network Coverage & Deployment Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific network architecture, segment structure, and
coverage assessment requirements.

Key customisation areas:
1. Network segment definitions (match your actual network topology)
2. Device categories and types (adapt to your environment)
3. Coverage verification methods (specific to your monitoring tools)
4. Exemption criteria and approval workflow (match your processes)
5. Performance validation thresholds (based on your requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.23 Web Filtering Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for verifying
web filtering coverage across all network segments, devices, and user groups.

**Purpose:**
Validates that web filtering controls are deployed comprehensively across the
organisation's network infrastructure, identifying coverage gaps and tracking
exemptions per ISO 27001:2022 Control A.8.23 requirements.

**Assessment Scope:**
- Complete network segment inventory and classification
- Device-level coverage verification across all segments
- User group and endpoint device coverage mapping
- Exemption identification, justification, and tracking
- Coverage verification through technical validation
- Gap analysis and remediation planning
- Evidence collection for comprehensive coverage audit

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and scoring guidance
2. Network Segment Inventory - Complete network topology documentation
3. Coverage Matrix - Segment-by-segment filtering coverage verification
4. Gap Analysis - Uncovered segments and remediation requirements
5. Device Inventory - Device-level coverage tracking and validation
6. Exemption Register - Approved exemptions with business justification
7. Coverage Verification - Technical validation of deployed controls
8. Evidence Register - Audit evidence and verification documentation
9. Approval & Sign-Off - Network team and security stakeholder approval

**Key Features:**
- Network topology visualization and segment classification
- Automated coverage percentage calculations
- Gap identification with risk-based prioritization
- Device-level tracking with exemption management
- Technical verification checklist and validation
- Evidence linking for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of four domains covering web filtering controls alongside
Infrastructure (A.8.23.1), Policy Configuration (A.8.23.3), and Monitoring
(A.8.23.4). Results feed into the Summary Dashboard for consolidated oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a823_2_network_coverage.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a823_2_network_coverage.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a823_2_network_coverage.py --date 20250115

Output:
    File: ISMS_A_8_23_2_Network_Coverage_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Populate complete network segment inventory from network diagrams
    2. Validate filtering deployment status for each segment
    3. Document device-level coverage across all endpoints
    4. Register and justify any coverage exemptions
    5. Perform technical verification of deployed controls
    6. Conduct gap analysis and define remediation timeline
    7. Collect and link technical validation evidence
    8. Obtain network operations and security team approvals
    9. Review Summary Dashboard metrics and finalise reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.23
Assessment Domain:    2 of 4 (Network Coverage & Deployment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.23: Web Filtering Policy (Governance)
    - ISMS-IMP-A.8.23.2: Network Coverage Implementation Guide
    - ISMS-IMP-A.8.23.1: Filtering Infrastructure Assessment (Domain 1)
    - ISMS-IMP-A.8.23.3: Policy Configuration Assessment (Domain 3)
    - ISMS-IMP-A.8.23.4: Monitoring & Response Assessment (Domain 4)

Related Controls:
    - A.8.20: Networks Security (Network segmentation and controls)
    - A.8.22: Segregation of Networks (Network isolation requirements)
    - A.5.15: Access Control (User access to network resources)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full network coverage assessment framework
    - Supports segment-level and device-level verification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
Network coverage verification is a critical audit evidence point. Ensure
complete segment inventory and accurate coverage status reporting. Auditors
will validate that ALL network segments are either filtered or have documented
exemptions with approved business justification.

**Data Protection:**
Assessment workbooks contain detailed network topology and security control
deployment information. Handle in accordance with your organisation's data
classification policies (typically INTERNAL or CONFIDENTIAL).

**Maintenance:**
Review and update network coverage assessments:
- Quarterly: Network topology changes, new segments, device additions
- After major infrastructure changes: Datacenter migrations, cloud adoption
- Annually: Complete re-verification of all coverage claims

**Quality Assurance:**
Network coverage assessments should be validated by:
- Network Operations team (topology accuracy)
- Security Operations team (control deployment validation)
- Technical verification through scanning or monitoring tools

**Common Pitfalls:**
- Incomplete network segment inventory (shadow IT, temporary networks)
- Undocumented exemptions (assume they'll be found during audit)
- Stale coverage data (verify regularly, don't rely on initial assessment)
- Missing mobile/remote device coverage (VPN, cloud-based filtering)

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.23.2"
WORKBOOK_NAME = "Network Coverage Assessment"
CONTROL_ID = "A.8.23"
CONTROL_NAME = "Web Filtering"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================


def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard sheet for Network Coverage Assessment."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(cell, text, bg="003366", fg="FFFFFF", bold=True, size=11, center=True):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        if center:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    def _grey(cell, value=None, center=False):
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        if value is not None:
            cell.value = value
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _yellow(cell):
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border

    # --- ROW 1 ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for _bc in range(1, 8):
        ws.cell(row=1, column=_bc).border = border

    # --- ROW 2 ---
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------------------
    # TABLE 1
    # -------------------------------------------------------------------------
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border

    row += 1
    t1_hdrs = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    _d9_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, h in enumerate(t1_hdrs, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="000000")
        cell.fill = _d9_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    start_row = row

    area_configs = [
        {
            "label": "Network Segment Inventory",
            "comp":    "=COUNTIF(\'Network Segment Inventory\'!J5:J55,\"\u2705 Protected\")",
            "partial": "=COUNTIF(\'Network Segment Inventory\'!J5:J55,\"\u26a0\ufe0f Partial\")",
            "noncomp": "=COUNTIF(\'Network Segment Inventory\'!J5:J55,\"\u274c Unprotected\")",
            "na":      "=COUNTIF(\'Network Segment Inventory\'!J5:J55,\"\u27f3 Planned\")"
                       "+COUNTIF(\'Network Segment Inventory\'!J5:J55,\"\u2298 Exempt\")",
        },
        {
            "label": "Coverage Matrix",
            "comp":    "=COUNTIF(\'Coverage Matrix\'!G5:G54,\"\u2705 Protected\")",
            "partial": "=COUNTIF(\'Coverage Matrix\'!G5:G54,\"\u26a0\ufe0f Partial\")",
            "noncomp": "=COUNTIF(\'Coverage Matrix\'!G5:G54,\"\u274c Unprotected\")",
            "na":      "=COUNTIF(\'Coverage Matrix\'!G5:G54,\"\u27f3 Planned\")"
                       "+COUNTIF(\'Coverage Matrix\'!G5:G54,\"\u2298 Exempt\")",
        },
        {
            "label": "Open Coverage Gaps",
            "comp":    "=COUNTIF(\'Gap Analysis\'!K5:K55,\"\u2705 Resolved\")"
                       "+COUNTIF(\'Gap Analysis\'!K5:K55,\"\u2298 Exempt\")",
            "partial": "=COUNTIF(\'Gap Analysis\'!K5:K55,\"\u27f3 In Progress\")",
            "noncomp": "=COUNTIF(\'Gap Analysis\'!K5:K55,\"\u25cb Open\")",
            "na":      "=COUNTIF(\'Gap Analysis\'!K5:K55,\"\u26a0\ufe0f Accepted Risk\")",
        },
    ]

    _blue_font = Font(color="000000")
    _blue_bold = Font(bold=True)
    _center_align = Alignment(horizontal="center", vertical="center")
    for cfg in area_configs:
        ws.cell(row=row, column=1, value=cfg["label"]).border = border
        ws.cell(row=row, column=2, value=f"=C{row}+D{row}+E{row}+F{row}").border = border
        ws.cell(row=row, column=2).font = _blue_font
        ws.cell(row=row, column=2).alignment = _center_align
        for col_idx, key in enumerate(["comp", "partial", "noncomp", "na"], 3):
            cell = ws.cell(row=row, column=col_idx, value=cfg[key])
            cell.font = _blue_font
            cell.border = border
            cell.alignment = _center_align
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.font = _blue_bold
        cell_g.border = border
        cell_g.alignment = _center_align
        cell_g.number_format = "0.0%"
        row += 1

    # TOTAL row
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = chr(64 + col_idx)
        cell = ws.cell(row=row, column=col_idx)
        cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{row - 1})"
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.border = border
        cell.alignment = _center_align
    cell_g = ws.cell(row=row, column=7)
    cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
    cell_g.font = Font(bold=True)
    cell_g.fill = PatternFill("solid", fgColor="D9D9D9")
    cell_g.border = border
    cell_g.alignment = _center_align
    cell_g.number_format = "0.0%"

    row += 3

    # -------------------------------------------------------------------------
    # TABLE 2: KEY METRICS
    # -------------------------------------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 2: KEY METRICS")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border
    row += 1
    _hdr(ws.cell(row=row, column=1), "Metric", bg="D9D9D9", fg="000000", size=10)
    ws.merge_cells(f"B{row}:G{row}")
    _hdr(ws.cell(row=row, column=2), "Count", bg="D9D9D9", fg="000000", size=10, center=True)
    for _c in range(3, 8):
        ws.cell(row=row, column=_c).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.cell(row=row, column=_c).border = border
    row += 1

    t2_metrics = [
        ("Segments ✅ Protected", "=COUNTIF('Network Segment Inventory'!J5:J55,\"✅ Protected\")"),
        ("Segments ❌ Unprotected", "=COUNTIF('Network Segment Inventory'!J5:J55,\"❌ Unprotected\")"),
        ("Coverage ✅ Protected", "=COUNTIF('Coverage Matrix'!G5:G54,\"✅ Protected\")"),
        ("Coverage ❌ Unprotected", "=COUNTIF('Coverage Matrix'!G5:G54,\"❌ Unprotected\")"),
        ("Coverage ⊘ Exempt / ⟳ Planned", "=COUNTIF('Coverage Matrix'!G5:G54,\"⊘ Exempt\")+COUNTIF('Coverage Matrix'!G5:G54,\"⟳ Planned\")"),
        ("Coverage Gaps ○ Open", "=COUNTIF('Gap Analysis'!K5:K55,\"○ Open\")"),
        ("Coverage Gaps ⟳ In Progress", "=COUNTIF('Gap Analysis'!K5:K55,\"⟳ In Progress\")"),
        ("Coverage Gaps ✅ Resolved / ⊘ Exempt", "=COUNTIF('Gap Analysis'!K5:K55,\"✅ Resolved\")+COUNTIF('Gap Analysis'!K5:K55,\"⊘ Exempt\")"),
    ]

    for _t2_label, _t2_formula in t2_metrics:
        _cell_a = ws.cell(row=row, column=1)
        _cell_a.value = _t2_label
        _cell_a.font = Font(name="Calibri")
        _cell_a.border = border
        ws.merge_cells(f"B{row}:G{row}")
        for _c in range(2, 8):
            ws.cell(row=row, column=_c).border = border
        ws.cell(row=row, column=2).value = _t2_formula
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).number_format = '0'
        row += 1

    row += 2

    # -------------------------------------------------------------------------
    # TABLE 3: CRITICAL FINDINGS & OPEN GAPS
    # -------------------------------------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 3: CRITICAL FINDINGS & OPEN GAPS", bg="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border
    row += 1
    t3_hdrs = ["Gap ID", "Segment Name", "Gap Description", "Risk Level",
               "Owner", "Target Date", "Status"]
    for col_idx, h in enumerate(t3_hdrs, 1):
        _hdr(ws.cell(row=row, column=col_idx), h, size=10)
    row += 1

    # TABLE 3: INDEX/SMALL/IF — auto-pull rows where Gap Analysis col K = "\u25cb Open"
    # Column mapping: Gap ID(A), Segment Name(C), Gap Description(E),
    #                 Risk Level(F), Owner(I), Target Date(J), Status(K)
    _gap = "Gap Analysis"
    _gcols = ["A", "C", "E", "F", "I", "J", "K"]
    for k in range(1, 11):
        for tbl_col, gc in enumerate(_gcols, 1):
            formula = (
                f"=IFERROR(INDEX(\'{_gap}\'!{gc}$5:{gc}$55,"
                f"SMALL(IF(\'{_gap}\'!K$5:K$55=\"\u25cb Open\"," 
                f"ROW(\'{_gap}\'!K$5:K$55)-ROW(\'{_gap}\'!K$5)+1),{k})),\"\")"
            )
            c = ws.cell(row=row, column=tbl_col)
            c.value = formula
            _yellow(c)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.freeze_panes = "A4"


def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.23.2 specification
    sheets = [
        "Instructions & Legend",
        "Network Segment Inventory",
        "Coverage Matrix",
        "Gap Analysis",
        "Device Inventory",
        "Exemption Register",
        "Coverage Verification",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Returns style TEMPLATES (dictionaries) to avoid shared object warnings.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_protected": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_unprotected": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        },
        "status_exempt": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        },
        "risk_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "risk_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "risk_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "risk_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell. Creates NEW objects to avoid shared warnings."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_restricted': DataValidation(
            type="list",
            formula1='"Yes,No,Restricted"',
            allow_blank=False
        ),
        'segment_type': DataValidation(
            type="list",
            formula1='"On-Premises LAN,WLAN,VPN,Cloud Endpoints,Guest,DMZ,Branch,Mobile,IoT/OT,Dev/Test,Other"',
            allow_blank=False
        ),
        'filtering_required': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'filtering_status': DataValidation(
            type="list",
            formula1='"\u2705 Protected,\u26A0\uFE0F Partial,\u274C Unprotected,\u27F3 Planned,\u2298 Exempt"',
            allow_blank=False
        ),
        'device_type': DataValidation(
            type="list",
            formula1='"Laptop,Desktop,Smartphone,Tablet,Virtual Desktop,Other"',
            allow_blank=False
        ),
        'agent_based': DataValidation(
            type="list",
            formula1='"Yes (endpoint agent),No (network-based),Hybrid"',
            allow_blank=False
        ),
        'exemption_type': DataValidation(
            type="list",
            formula1='"Full Exemption,Partial Exemption,Temporary Exemption,Category Exemption,Technical Exemption"',
            allow_blank=False
        ),
        'exemption_status': DataValidation(
            type="list",
            formula1='"\u2705 Active,\u274c Expired,\u26d4 Revoked,\u23f3 Under Review"',
            allow_blank=False
        ),
        'test_method': DataValidation(
            type="list",
            formula1='"Manual Browse Test,Automated Scan,Curl/wget Test,Browser Extension,Vendor Tool,Penetration Test,Other"',
            allow_blank=False
        ),
        'test_result': DataValidation(
            type="list",
            formula1='"Pass,Fail,Partial,Inconclusive"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1='"\u2705 Verified,\u274C Failed,\u26A0\uFE0F Needs Retest"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"○ Open,⟳ In Progress,✅ Resolved,⚠️ Accepted Risk,⊘ Exempt"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Network Diagram,VLAN Config,Firewall Policy,VPN Config,WiFi Config,Cloud Dashboard,NAC Policy,Exemption Approval,Test Results,Monitoring Report,Change Record,Other"',
            allow_blank=False
        ),
        'evidence_verification': DataValidation(
            type="list",
            formula1='"Verified,Pending,Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Final,Requires Remediation,Re-assessment Required"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected"',
            allow_blank=False
        ),
        'recommendation': DataValidation(
            type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False
        ),
    }

    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Network Segment Inventory for every segment in your environment.', '2. Document your specific network architecture (office LANs, WLANs, VPN, cloud endpoints, etc.).', '3. Map filtering solutions to network segments in the Coverage Matrix.', '4. Calculate coverage percentages and identify gaps in the Gap Analysis sheet.', '5. Document any devices or segments exempt from filtering in the Exemption Register.', '6. Perform technical verification tests and record results in Coverage Verification.', '7. Track all supporting evidence in the Evidence Register.', '8. Obtain network security and CISO approval via the Approval Sign-Off sheet.', '9. Update quarterly or after any network topology changes.', '10. Coordinate with ISMS-IMP-A.8.23.1 (Infrastructure Assessment) for cross-reference.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 24

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_network_segment_inventory(ws, styles):
    """Create Network Segment Inventory sheet - foundation of coverage analysis."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:O1")
    ws["A1"] = "NETWORK SEGMENT INVENTORY"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:O2")
    ws["A2"] = "Document all network segments requiring web filtering protection"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Segment ID", 15),
        ("B", "Segment Name", 30),
        ("C", "Segment Type", 25),
        ("D", "Location/Site", 25),
        ("E", "Subnet/VLAN", 20),
        ("F", "User Count", 12),
        ("G", "Device Count", 12),
        ("H", "Internet Access?", 15),
        ("I", "Filtering Required?", 15),
        ("J", "Filtering Status", 18),
        ("K", "Filtering Solution(s)", 30),
        ("L", "Coverage %", 12),
        ("M", "Bypass Methods", 25),
        ("N", "Exemption ID", 15),
        ("O", "Evidence", 35),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Data rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1

    # Sample row with SEG-001
    ws[f"A{row}"] = "SEG-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border

    validations['segment_type'].add(ws[f"C{row}"])
    validations['yes_no_restricted'].add(ws[f"H{row}"])
    validations['filtering_required'].add(ws[f"I{row}"])
    validations['filtering_status'].add(ws[f"J{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border
        row += 1

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 5: COVERAGE_MATRIX SHEET
# ============================================================================

def create_coverage_matrix(ws, styles):
    """Create Coverage Matrix sheet - cross-reference segments with solutions."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "NETWORK COVERAGE MATRIX"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Map filtering solutions to network segments"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Network Segment", "Solution 1", "Solution 2", "Solution 3", "Solution 4", "Total Coverage", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows (50 segments matching inventory)
    row += 1
    for i in range(1, 51):
        # Segment reference — only first row has example text; rest left blank
        if i == 1:
            ws[f"A{row}"] = "SEG-001 [Segment Name]"
            ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Solution columns (B-E) - customer fills in coverage %
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Total Coverage (formula - sum of solution percentages)
        ws[f"F{row}"] = f'=SUM(B{row}:E{row})'
        ws[f"F{row}"].number_format = '0"%"'
        ws[f"F{row}"].font = Font(bold=True, color="0000FF")

        # Status (conditional based on total coverage)
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = border
        validations['filtering_status'].add(ws[f"G{row}"])
        
        row += 1

    # Summary section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COVERAGE SUMMARY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    summary_items = [
        ("Total segments assessed:", f'=COUNTA(A5:A54)'),
        ("Fully protected (100%):", f'=COUNTIF(F5:F54,">=100")'),
        ("Partially protected (<100%):", f'=COUNTIFS(F5:F54,">0",F5:F54,"<100")'),
        ("Unprotected (0%):", f'=COUNTIF(F5:F54,0)'),
        ("Overall Coverage Score:", f'=AVERAGE(F5:F54)&"%"'),
    ]

    for label, formula in summary_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, size=11, color="0000FF")
        row += 1

    # Coverage Heatmap
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "COVERAGE HEATMAP"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Coverage Level"
    ws[f"B{row}"] = "Segment Count"
    ws[f"C{row}"] = "% of Total"
    ws[f"D{row}"] = "Status"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    heatmap_start_row = row
    heatmap_end_row = row + 4  # 5 coverage levels
    heatmap_data = [
        ("100% Coverage", f'=COUNTIF(F5:F54,100)', "\u2705 Good", "C6EFCE"),
        ("75-99% Coverage", f'=COUNTIFS(F5:F54,">=75",F5:F54,"<100")', "\u26A0\uFE0F Needs Attention", "FFEB9C"),
        ("50-74% Coverage", f'=COUNTIFS(F5:F54,">=50",F5:F54,"<75")', "\u26A0\uFE0F Risk", "FFEB9C"),
        ("<50% Coverage", f'=COUNTIFS(F5:F54,">0",F5:F54,"<50")', "\u274C Critical", "FFC7CE"),
        ("0% Coverage", f'=COUNTIF(F5:F54,0)', "\u274C Urgent", "FFC7CE"),
    ]

    for level, formula, status, color in heatmap_data:
        ws[f"A{row}"] = level
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/SUM(B{heatmap_start_row}:B{heatmap_end_row})*100)&"%"'
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        ws[f"D{row}"] = status
        ws[f"D{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 6: GAP_IDENTIFICATION SHEET
# ============================================================================

def create_gap_identification(ws, styles):
    """Create Gap Analysis sheet - track coverage gaps and remediation."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "COVERAGE GAPS & REMEDIATION TRACKING"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    ws["A2"] = "Identify unprotected/partially protected segments and plan remediation"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        "Gap ID", "Segment ID", "Segment Name", "Current Coverage", "Gap Description",
        "Risk Level", "Impact", "Remediation Plan", "Owner", "Target Date", "Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Gap tracking rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1
    gap_start_row = row

    # Sample row with GAP-001
    _gap_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws[f"A{row}"] = "GAP-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)
    ws[f"A{row}"].fill = _gap_grey
    ws[f"A{row}"].border = border

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws[f"{col}{row}"].fill = _gap_grey
        ws[f"{col}{row}"].border = border

    validations['risk_level'].add(ws[f"F{row}"])
    validations['gap_status'].add(ws[f"K{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"A{row}"].border = border
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border
        row += 1

    gap_end_row = row - 1

    # Gap Summary Metrics
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "GAP SUMMARY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Risk Level"
    ws[f"B{row}"] = "Gap Count"
    ws[f"C{row}"] = "% of Total"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    risk_levels = [
        ("Critical", "C00000", "FFFFFF"),
        ("High", "FFC7CE", "000000"),
        ("Medium", "FFEB9C", "000000"),
        ("Low", "C6EFCE", "000000"),
    ]

    for risk, bg_color, fg_color in risk_levels:
        ws[f"A{row}"] = risk
        ws[f"A{row}"].font = Font(bold=True, color=fg_color)
        ws[f"A{row}"].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        ws[f"B{row}"] = f'=COUNTIF(F{gap_start_row}:F{gap_end_row},"{risk}")'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/COUNTA(F{gap_start_row}:F{gap_end_row})*100)&"%"'
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        
        row += 1

    # Overall summary
    row += 1
    overall_items = [
        ("Total gaps identified:", f'=COUNTA(B{gap_start_row}:B{gap_end_row})'),
        ("Gaps resolved:", f'=COUNTIF(K{gap_start_row}:K{gap_end_row},"\u2705 Resolved")'),
        ("Gaps remaining:", f'=COUNTA(B{gap_start_row}:B{gap_end_row})-COUNTIF(K{gap_start_row}:K{gap_end_row},"\u2705 Resolved")'),
    ]

    for label, formula in overall_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, size=11, color="C00000" if "remaining" in label else "0000FF")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15

    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7: DEVICE_INVENTORY SHEET
# ============================================================================

def create_device_inventory(ws, styles):
    """Create Device Inventory sheet - endpoint-level tracking."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = "DEVICE-LEVEL FILTERING INVENTORY"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"] = "Track endpoint protection for mobile/remote devices"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Device ID", 12),
        ("B", "Device Type", 15),
        ("C", "OS", 15),
        ("D", "User/Owner", 25),
        ("E", "Primary Network", 25),
        ("F", "Filtering Solution", 25),
        ("G", "Agent-Based?", 18),
        ("H", "Status", 15),
        ("I", "Last Verified", 15),
        ("J", "Evidence", 35),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Device rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1

    # Sample row with DEV-001
    ws[f"A{row}"] = "DEV-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border

    validations['device_type'].add(ws[f"B{row}"])
    validations['agent_based'].add(ws[f"G{row}"])
    validations['filtering_status'].add(ws[f"H{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border
        row += 1

    # Device Summary
    row += 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "DEVICE SUMMARY BY TYPE"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Device Type"
    ws[f"B{row}"] = "Total"
    ws[f"C{row}"] = "Protected"
    ws[f"D{row}"] = "Unprotected"
    ws[f"E{row}"] = "Coverage %"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    device_types = ["Laptop", "Desktop", "Smartphone", "Tablet", "Virtual Desktop", "Other"]
    
    for dev_type in device_types:
        ws[f"A{row}"] = dev_type
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f'=COUNTIF(B5:B55,"{dev_type}")'
        ws[f"C{row}"] = f'=COUNTIFS(B5:B55,"{dev_type}",H5:H55,"\u2705 Protected")'
        ws[f"D{row}"] = f'=COUNTIFS(B5:B55,"{dev_type}",H5:H55,"\u274C Unprotected")'
        ws[f"E{row}"] = f'=IF(B{row}=0,0,C{row}/B{row}*100)&"%"'
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: EXEMPTION_REGISTER SHEET
# ============================================================================

def create_exemption_register(ws, styles):
    """Create Exemption Register sheet - approved exceptions documentation."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "FILTERING EXEMPTION REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    ws["A2"] = "Document approved exceptions to web filtering requirements"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        "Exemption ID", "Segment/Device", "Exemption Type", "Business Justification",
        "Risk Assessment", "Compensating Controls", "Approved By", "Approval Date",
        "Review Date", "Status", "Evidence"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Exemption rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1

    # Sample row with EXM-001
    ws[f"A{row}"] = "EXM-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border

    validations['exemption_type'].add(ws[f"C{row}"])
    validations['exemption_status'].add(ws[f"J{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border
        row += 1

    # Exemption Summary
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "EXEMPTION SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Exemption Type"
    ws[f"B{row}"] = "Active Count"
    ws[f"C{row}"] = "Expired/Overdue"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    exemption_types = [
        "Full Exemption",
        "Partial Exemption",
        "Temporary Exemption",
        "Category Exemption",
        "Technical Exemption",
    ]

    for exm_type in exemption_types:
        ws[f"A{row}"] = exm_type
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f'=COUNTIFS(C5:C24,"{exm_type}",J5:J24,"\u2705 Active")'
        ws[f"C{row}"] = f'=COUNTIFS(C5:C24,"{exm_type}",J5:J24,"\u274c Expired")'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"].font = Font(bold=True, color="C00000")
        row += 1

    # Alert section
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "\u26A0\uFE0F ALERT: Exemptions >12 months old require re-approval | >10% of segments exempt = escalate to CISO"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 30

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 9: COVERAGE_VERIFICATION SHEET
# ============================================================================

def create_coverage_verification(ws, styles):
    """Create Coverage Verification sheet - testing documentation."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = "COVERAGE VERIFICATION LOG"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"] = "Test and verify filtering effectiveness per network segment"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Verification ID", 15),
        ("B", "Segment ID", 12),
        ("C", "Segment Name", 25),
        ("D", "Test Date", 15),
        ("E", "Tester", 20),
        ("F", "Test Method", 20),
        ("G", "Test Results", 30),
        ("H", "Issues Found", 35),
        ("I", "Status", 15),
        ("J", "Next Test Date", 15),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Verification rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1

    # Sample row with VER-001
    ws[f"A{row}"] = "VER-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    # Segment ID, Segment Name, Tester, Test Results, Issues
    for col in ["B", "C", "E", "G", "H"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border

    # Test Date
    ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"D{row}"].border = border

    # Test Method dropdown
    ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"F{row}"].border = border
    validations['test_method'].add(ws[f"F{row}"])

    # Status dropdown
    ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"I{row}"].border = border
    validations['verification_status'].add(ws[f"I{row}"])

    # Next Test Date (auto-calculate +90 days)
    ws[f"J{row}"] = f'=IF(D{row}="","",D{row}+90)'
    ws[f"J{row}"].number_format = 'DD.MM.YYYY'

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        # Segment ID, Segment Name, Tester, Test Results, Issues
        for col in ["B", "C", "E", "G", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

        # Test Date
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = border

        # Test Method dropdown
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = border
        validations['test_method'].add(ws[f"F{row}"])

        # Status dropdown
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"I{row}"].border = border
        validations['verification_status'].add(ws[f"I{row}"])

        # Next Test Date (auto-calculate +90 days)
        ws[f"J{row}"] = f'=IF(D{row}="","",D{row}+90)'
        ws[f"J{row}"].number_format = 'DD.MM.YYYY'

        row += 1

    # Test Checklist Template
    row += 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "TEST CHECKLIST (Verify for each segment)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "☐ Malicious URL blocked (test with EICAR or known bad domains)",
        "☐ Phishing site blocked (test with PhishTank samples)",
        "☐ Known malware download blocked",
        "☐ HTTPS sites properly filtered (if HTTPS inspection enabled)",
        "☐ Logging working (blocked attempts logged)",
        "☐ Bypass methods ineffective (proxy bypass, VPN bypass, etc.)",
        "☐ User experience acceptable (latency, false positives)",
        "☐ Exemptions working as intended",
    ]

    row += 1
    for item in checklist_items:
        ws[f"A{row}"] = item
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: EVIDENCE_REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (italic, NOT blue banner) ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers (003366, white text) ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,'
                 'Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Web Filtering Network Coverage",
        3: "Configuration file",
        4: "Network segment inventory and filtering coverage evidence",
        5: "/evidence/A.8.23/",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None  # Empty — users choose their own evidence IDs
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL_SIGN_OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # ── Row 2: Control reference (italic, 003366 text) ──
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # ── Row 3: ASSESSMENT SUMMARY banner ──
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # ── Summary fields (Rows 4-8) ──
    # (label, value, ffffcc_fill)
    summary_fields = [
        ("Document:",                  f"{DOCUMENT_ID} - {WORKBOOK_NAME}", False),
        ("Assessment Period:",         "",                                   True),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9",          True),
        ("Assessment Status:",         "",                                   True),
        ("Assessed By:",               "",                                   True),
    ]

    row = 4
    status_row_for_dv = None
    for label, value, yellow in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for col in range(2, 6):
            if yellow:
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        if "Assessment Status" in label:
            status_row_for_dv = row
        row += 1

    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f"B{status_row_for_dv}")

    # ── Approver sections ──
    row += 2  # gap before first approver

    def _approver(start_row, title, color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = border
            ws.merge_cells(f"B{r}:E{r}")
            for col in range(2, 6):
                ws.cell(row=r, column=col).fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=r, column=col).border = border
            r += 1
        return r + 1  # next row after gap

    row = _approver(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _approver(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")
    row = _approver(row, "APPROVED BY (CISO)", "003366")

    # ── Final Decision ──
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f"B{row}")

    # ── Next Review Details ──
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        row += 1

    # ── Column widths & freeze ──
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Network coverage is where the rubber meets the road. You can have the best
    filtering technology in the world, but if it only covers 60% of your network,
    you have a 40% gap for attackers to exploit.
    
    This assessment answers: WHERE is filtering applied, and WHERE are the gaps?
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.23.2 - Network Coverage Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.23: Web Filtering")
    logger.info("=" * 78)
    logger.info("\n[*] Network Coverage: Verify filtering across ALL network segments")
    logger.info("[*] Topology-Agnostic: Works with ANY network architecture")
    logger.info("[*] Gap Analysis: Find unprotected segments")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("\u2705 Workbook created with 10 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")
    
    logger.info("  [1/9] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  \u2705 Instructions complete")

    logger.info("  [2/9] Creating Network Segment Inventory...")
    create_network_segment_inventory(wb["Network Segment Inventory"], styles)
    logger.info("  \u2705 Segment inventory complete (50 segments)")

    logger.info("  [3/9] Creating Coverage Matrix...")
    create_coverage_matrix(wb["Coverage Matrix"], styles)
    logger.info("  \u2705 Coverage matrix complete (cross-reference map)")

    logger.info("  [4/9] Creating Gap Analysis...")
    create_gap_identification(wb["Gap Analysis"], styles)
    logger.info("  \u2705 Gap tracking complete (30 gap rows)")

    logger.info("  [5/9] Creating Device Inventory...")
    create_device_inventory(wb["Device Inventory"], styles)
    logger.info("  \u2705 Device inventory complete (100 devices)")

    logger.info("  [6/9] Creating Exemption Register...")
    create_exemption_register(wb["Exemption Register"], styles)
    logger.info("  \u2705 Exemption register complete (20 exemptions)")

    logger.info("  [7/9] Creating Coverage Verification...")
    create_coverage_verification(wb["Coverage Verification"], styles)
    logger.info("  \u2705 Verification log complete (50 tests)")


    logger.info("  Creating Summary Dashboard...")


    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)


    logger.info("  \u2705 Summary Dashboard complete")


    logger.info("  [8/9] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  \u2705 Evidence register complete (100 evidence rows)")

    logger.info("  [9/9] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  \u2705 Approval workflow complete")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.23.2_Network_Coverage_{datetime.now().strftime('%Y%m%d')}.xlsx"

    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"\u2705 SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"\u274C ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\nCore Assessment:")
    logger.info("  \u2022 Instructions & Legend (usage guidance + segment types)")
    logger.info("  \u2022 Network Segment Inventory (50 segments - foundation)")
    logger.info("  \u2022 Coverage Matrix (cross-reference solutions to segments)")
    logger.info("  \u2022 Gap Analysis (30 gap tracking rows)")
    logger.info("\nSupplementary Tracking:")
    logger.info("  \u2022 Device Inventory (100 endpoint devices)")
    logger.info("  \u2022 Exemption Register (20 approved exemptions)")
    logger.info("  \u2022 Coverage Verification (50 test entries)")
    logger.info("\nGovernance:")
    logger.info("  \u2022 Evidence Register (100 evidence entries)")
    logger.info("  \u2022 Approval Sign-Off (3-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info("ASSESSMENT CAPABILITIES:")
    logger.info("  \u2022 50 network segments documented")
    logger.info("  \u2022 Cross-reference with 4 filtering solutions")
    logger.info("  \u2022 100 endpoint devices tracked")
    logger.info("  \u2022 30 gap identification/remediation rows")
    logger.info("  \u2022 20 exemption approvals")
    logger.info("  \u2022 50 verification test entries")
    logger.info("  \u2022 Auto-calculated coverage percentages")
    logger.info("  \u2022 Coverage heatmap analysis")
    logger.info("\n" + "─" * 78)
    logger.info("KEY FEATURES:")
    logger.info("  \u2705 Topology-agnostic (map ANY network architecture)")
    logger.info("  \u2705 Comprehensive segment coverage tracking")
    logger.info("  \u2705 Automated coverage percentage calculations")
    logger.info("  \u2705 Gap prioritization by risk level")
    logger.info("  \u2705 Exemption management with approval workflow")
    logger.info("  \u2705 Verification testing documentation")
    logger.info("  \u2705 Multi-level approval process")
    logger.info("\n" + "=" * 78)
    logger.info("NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Document ALL network segments in Network Segment Inventory")
    logger.info("  3. Map solutions to segments in Coverage Matrix")
    logger.info("  4. Calculate coverage percentages (auto-calculated)")
    logger.info("  5. Identify gaps in Gap Analysis")
    logger.info("  6. Document exemptions if needed")
    logger.info("  7. Verify coverage with testing (Coverage Verification)")
    logger.info("  8. Obtain approvals via Approval Sign-Off")
    logger.info("\nPRO TIP:")
    logger.info("  Coverage theater says 'we have web filtering.'")
    logger.info("  Systems engineering says 'we have 97.3% coverage across 47 segments")
    logger.info("  with 3 approved exemptions and quarterly verification testing.'")
    logger.info("\nCOVERAGE FORMULA:")
    logger.info("  Network Coverage Score = Σ(Protected Segments) / Total Segments * 100%")
    logger.info("\n" + "=" * 78)
    logger.info('\n"Cargo cult ISMS: We have filtering. Somewhere. Maybe."')
    logger.info('"Systems Engineering: Verified 97.3% coverage with evidence."')
    logger.info("\nEvidence > Theater")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
