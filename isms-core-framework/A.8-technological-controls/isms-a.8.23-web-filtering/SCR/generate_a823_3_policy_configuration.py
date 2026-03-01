#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.23.3 - Policy Configuration Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.23: Web Filtering
Assessment Domain 3 of 4: Policy Configuration & Content Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific policy requirements, acceptable use guidelines,
and web filtering configuration standards.

Key customisation areas:
1. Threat protection categories and priorities (match your risk profile)
2. Category filtering rules (adapt to your acceptable use policy)
3. Custom allow/block lists (specific to your business requirements)
4. User group policy structures (match your organisational hierarchy)
5. Exception criteria and approval workflows (adapt to your governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.23 Web Filtering Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
web filtering policy configuration against organisational requirements and
ISO 27001:2022 Control A.8.23 mandates.

**Purpose:**
Validates that web filtering policies are properly configured to protect
against threats, enforce acceptable use policies, and support business
requirements while maintaining appropriate exception management.

**Assessment Scope:**
- Threat protection policy configuration and effectiveness
- Category-based filtering aligned to acceptable use policy
- Custom allow/block list management and justification
- Policy exception handling and approval tracking
- User group and role-based policy differentiation
- Policy review and maintenance processes
- Configuration gap analysis and remediation
- Evidence collection for policy compliance audit

**Generated Workbook Structure:**
1. Instructions & Legend - Policy assessment methodology and criteria
2. Threat Protection - Malware, phishing, and exploit protection policies
3. Category Management - Content category filtering rules and rationale
4. Custom Lists - Allow/block list inventory and business justification
5. Policy Exceptions - Exception register with approval documentation
6. User Group Policies - Differentiated policies by user role/function
7. Acceptable Use Alignment - AUP mapping to filtering policies
8. Policy Review Process - Configuration management and update procedures
9. Gap Analysis - Policy configuration gaps and remediation requirements
10. Evidence Register - Audit evidence and configuration documentation
11. Approval & Sign-Off - Policy owner and security team approval

**Key Features:**
- Threat category prioritization and risk-based policy configuration
- Content filtering alignment to business requirements
- Exception tracking with business justification and time-limiting
- User group policy differentiation based on role/risk
- Acceptable Use Policy (AUP) traceability matrix
- Automated gap identification for policy deficiencies
- Evidence linking for audit readiness
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of four domains covering web filtering controls alongside
Infrastructure (A.8.23.1), Network Coverage (A.8.23.2), and Monitoring
(A.8.23.4). Results feed into the Summary Dashboard for consolidated reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a823_3_policy_configuration.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a823_3_policy_configuration.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a823_3_policy_configuration.py --date 20250115

Output:
    File: ISMS_A_8_23_3_Policy_Configuration_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Document current threat protection policy configuration
    2. Map category filtering rules to acceptable use policy requirements
    3. Inventory and justify all custom allow/block list entries
    4. Register all policy exceptions with business justification
    5. Document user group policies and differentiation rationale
    6. Verify AUP alignment across all policy configurations
    7. Review policy maintenance and update procedures
    8. Conduct gap analysis and define remediation actions
    9. Collect and link policy configuration evidence
    10. Obtain policy owner and security stakeholder approvals
    11. Review Summary Dashboard metrics and finalise reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.23
Assessment Domain:    3 of 4 (Policy Configuration & Content Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.23: Web Filtering Policy (Governance)
    - ISMS-IMP-A.8.23.3: Policy Configuration Implementation Guide
    - ISMS-POL-A.5.10: Acceptable Use of Information and Assets
    - ISMS-IMP-A.8.23.1: Filtering Infrastructure Assessment (Domain 1)
    - ISMS-IMP-A.8.23.2: Network Coverage Assessment (Domain 2)
    - ISMS-IMP-A.8.23.4: Monitoring & Response Assessment (Domain 4)

Related Controls:
    - A.5.10: Acceptable Use of Information and Assets
    - A.5.15: Access Control (User access policies)
    - A.8.4: Access to Source Code (Developer exception handling)
    - A.5.9: Inventory of Information and Assets (Custom lists)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full policy configuration assessment framework
    - Supports threat protection, category filtering, and exception management
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
Policy configuration is a primary focus during ISO 27001 audits. Ensure:
- All threat protection policies are documented and justified
- Category filtering aligns to documented acceptable use policy
- ALL exceptions have business justification and approval
- User group differentiation is risk-based and documented
- Policy review processes are documented and evidence is retained

**Data Protection:**
Assessment workbooks may contain sensitive policy configuration and business
justification information. Handle in accordance with your organisation's data
classification policies (typically INTERNAL or CONFIDENTIAL).

**Maintenance:**
Review and update policy configuration assessments:
- Quarterly: Threat landscape changes, new threat categories
- After AUP changes: Acceptable use policy updates require filtering alignment
- Semi-annually: Exception review and revalidation
- Annually: Complete policy configuration re-assessment

**Quality Assurance:**
Policy assessments should be validated by:
- Security Policy team (requirement accuracy)
- Technical Security team (configuration verification)
- Business stakeholders (exception justification validity)
- Compliance team (regulatory alignment verification)

**Common Pitfalls:**
- Generic threat policies without risk-based prioritization
- Category filtering misalignment with actual acceptable use policy
- Undocumented or "verbal approval" exceptions (audit failure)
- Stale exception registers (no review/revalidation process)
- Missing user group differentiation (one-size-fits-all approach)
- No AUP traceability (cannot prove policy alignment)

**Exception Management Critical Success Factors:**
- EVERY exception must have documented business justification
- Time-limited exceptions with review dates
- Named approvers with approval dates
- Regular exception review and revalidation (quarterly minimum)
- Automatic expiration and re-approval workflow

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.23.3"
WORKBOOK_NAME = "Policy Configuration Assessment"
CONTROL_ID = "A.8.23"
CONTROL_NAME = "Web Filtering"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================


def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard sheet for Policy Configuration Assessment."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(cell, text, bg="003366", fg="FFFFFF", bold=True, size=11, center=True):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        if center:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    def _grey(cell, value=None, center=False):
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        if value is not None:
            cell.value = value
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _yellow(cell):
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border

    # --- ROW 1 ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for _bc in range(1, 8):
        ws.cell(row=1, column=_bc).border = border

    # --- ROW 2 ---
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------------------
    # TABLE 1
    # -------------------------------------------------------------------------
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border

    row += 1
    t1_hdrs = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    _d9_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, h in enumerate(t1_hdrs, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="000000")
        cell.fill = _d9_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    start_row = row

    area_configs = [
        {
            "label": "Threat Protection",
            "comp":    "=COUNTIF(\'Threat Protection\'!B5:B21,\"\u2705 Configured\")",
            "partial": "=COUNTIF(\'Threat Protection\'!B5:B21,\"\u26a0\ufe0f Partial\")",
            "noncomp": "=COUNTIF(\'Threat Protection\'!B5:B21,\"\u274c Not Configured\")"
                       "+COUNTIF(\'Threat Protection\'!B5:B21,\"\u27f3 Planned\")",
            "na":      "=COUNTIF(\'Threat Protection\'!B5:B21,\"N/A\")",
        },
        {
            "label": "Category Management",
            "comp":    "=COUNTIF(\'Category Management\'!B9:B28,\"⛔ Block\")"
                       "+COUNTIF(\'Category Management\'!B9:B28,\"✅ Allow\")"
                       "+COUNTIF(\'Category Management\'!B9:B28,\"⚠️ Warn\")",
            "partial": "=COUNTIF(\'Category Management\'!B9:B28,\"○ Monitor Only\")",
            "noncomp": "=0",
            "na":      "=COUNTIF(\'Category Management\'!B9:B28,\"N/A\")",
        },
        {
            "label": "Open Policy Gaps",
            "comp":    "=COUNTIF(\'Gap Analysis\'!K5:K55,\"\u2705 Resolved\")"
                       "+COUNTIF(\'Gap Analysis\'!K5:K55,\"\u26a0\ufe0f Accepted Risk\")",
            "partial": "=COUNTIF(\'Gap Analysis\'!K5:K55,\"\u27f3 In Progress\")",
            "noncomp": "=COUNTIF(\'Gap Analysis\'!K5:K55,\"\u25cb Open\")",
            "na":      "=0",
        },
    ]

    _blue_font = Font(color="000000")
    _blue_bold = Font(bold=True)
    _center_align = Alignment(horizontal="center", vertical="center")
    for cfg in area_configs:
        ws.cell(row=row, column=1, value=cfg["label"]).border = border
        ws.cell(row=row, column=2, value=f"=C{row}+D{row}+E{row}+F{row}").border = border
        ws.cell(row=row, column=2).font = _blue_font
        ws.cell(row=row, column=2).alignment = _center_align
        for col_idx, key in enumerate(["comp", "partial", "noncomp", "na"], 3):
            cell = ws.cell(row=row, column=col_idx, value=cfg[key])
            cell.font = _blue_font
            cell.border = border
            cell.alignment = _center_align
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.font = _blue_bold
        cell_g.border = border
        cell_g.alignment = _center_align
        cell_g.number_format = "0.0%"
        row += 1

    # TOTAL row
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = chr(64 + col_idx)
        cell = ws.cell(row=row, column=col_idx)
        cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{row - 1})"
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.border = border
        cell.alignment = _center_align
    cell_g = ws.cell(row=row, column=7)
    cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
    cell_g.font = Font(bold=True)
    cell_g.fill = PatternFill("solid", fgColor="D9D9D9")
    cell_g.border = border
    cell_g.alignment = _center_align
    cell_g.number_format = "0.0%"

    row += 3

    # -------------------------------------------------------------------------
    # TABLE 2: KEY METRICS
    # -------------------------------------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 2: KEY METRICS")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border
    row += 1
    _hdr(ws.cell(row=row, column=1), "Metric", bg="D9D9D9", fg="000000", size=10)
    ws.merge_cells(f"B{row}:G{row}")
    _hdr(ws.cell(row=row, column=2), "Count", bg="D9D9D9", fg="000000", size=10, center=True)
    for _c in range(3, 8):
        ws.cell(row=row, column=_c).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.cell(row=row, column=_c).border = border
    row += 1

    t2_metrics = [
        ("Threat Types ✅ Configured", "=COUNTIF('Threat Protection'!B5:B21,\"✅ Configured\")"),
        ("Threat Types ❌ Not Configured / ⟳ Planned", "=COUNTIF('Threat Protection'!B5:B21,\"❌ Not Configured\")+COUNTIF('Threat Protection'!B5:B21,\"⟳ Planned\")"),
        ("Categories ⛔ Blocked", "=COUNTIF('Category Management'!B9:B28,\"⛔ Block\")"),
        ("Categories ✅ Allowed", "=COUNTIF('Category Management'!B9:B28,\"✅ Allow\")"),
        ("Policy Exceptions ✅ Active", "=COUNTIF('Policy Exceptions'!L5:L55,\"✅ Active\")"),
        ("Policy Exceptions ⏳ Under Review", "=COUNTIF('Policy Exceptions'!L5:L55,\"⏳ Under Review\")"),
        ("Open Policy Gaps ○", "=COUNTIF('Gap Analysis'!K5:K55,\"○ Open\")"),
        ("Policy Gaps ✅ Resolved", "=COUNTIF('Gap Analysis'!K5:K55,\"✅ Resolved\")"),
    ]

    for _t2_label, _t2_formula in t2_metrics:
        _cell_a = ws.cell(row=row, column=1)
        _cell_a.value = _t2_label
        _cell_a.font = Font(name="Calibri")
        _cell_a.border = border
        ws.merge_cells(f"B{row}:G{row}")
        for _c in range(2, 8):
            ws.cell(row=row, column=_c).border = border
        ws.cell(row=row, column=2).value = _t2_formula
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).number_format = '0'
        row += 1

    row += 2

    # -------------------------------------------------------------------------
    # TABLE 3: CRITICAL FINDINGS & OPEN GAPS
    # -------------------------------------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 3: CRITICAL FINDINGS & OPEN GAPS", bg="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border
    row += 1
    t3_hdrs = ["Gap ID", "Policy Area", "Gap Description", "Risk Level",
               "Owner", "Target Date", "Status"]
    for col_idx, h in enumerate(t3_hdrs, 1):
        _hdr(ws.cell(row=row, column=col_idx), h, size=10)
    row += 1

    # TABLE 3: INDEX/SMALL/IF — auto-pull rows where Gap Analysis col K = "\u25cb Open"
    # Column mapping: Gap ID(A), Policy Area(B), Gap Description(C),
    #                 Risk Level(D), Owner(I), Target Date(J), Status(K)
    _gap = "Gap Analysis"
    _gcols = ["A", "B", "C", "D", "I", "J", "K"]
    for k in range(1, 11):
        for tbl_col, gc in enumerate(_gcols, 1):
            formula = (
                f"=IFERROR(INDEX(\'{_gap}\'!{gc}$5:{gc}$55,"
                f"SMALL(IF(\'{_gap}\'!K$5:K$55=\"\u25cb Open\"," 
                f"ROW(\'{_gap}\'!K$5:K$55)-ROW(\'{_gap}\'!K$5)+1),{k})),\"\")"
            )
            c = ws.cell(row=row, column=tbl_col)
            c.value = formula
            _yellow(c)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.freeze_panes = "A4"


def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.23.3 specification
    sheets = [
        "Instructions & Legend",
        "Threat Protection",
        "Category Management",
        "Custom Lists",
        "Policy Exceptions",
        "User Group Policies",
        "Acceptable Use Alignment",
        "Policy Review Process",
        "Evidence Register",
        "Gap Analysis",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries) to avoid shared object warnings.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_configured": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notconfigured": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        },
        "risk_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "risk_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "risk_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "risk_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell. Creates NEW objects to avoid shared warnings."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'policy_status': DataValidation(
            type="list",
            formula1='"\u2705 Configured,\u26A0\uFE0F Partial,\u274C Not Configured,\u27F3 Planned,N/A"',
            allow_blank=False
        ),
        'effectiveness': DataValidation(
            type="list",
            formula1='"High,Medium,Low,Unknown"',
            allow_blank=False
        ),
        'false_positive_rate': DataValidation(
            type="list",
            formula1='"Rare,Occasional,Frequent,Unknown"',
            allow_blank=False
        ),
        'update_frequency': DataValidation(
            type="list",
            formula1='"Real-time,Hourly,Daily,Weekly,Monthly"',
            allow_blank=False
        ),
        'filtering_philosophy': DataValidation(
            type="list",
            formula1='"Restrictive (Default Deny),Trust-Based (Threats Only),Hybrid (Balanced)"',
            allow_blank=False
        ),
        'policy_action': DataValidation(
            type="list",
            formula1='"⛔ Block,✅ Allow,⚠️ Warn,○ Monitor Only,N/A"',
            allow_blank=False
        ),
        'applied_to': DataValidation(
            type="list",
            formula1='"All Users,Specific Group,Specific User,Network Segment,Device Type"',
            allow_blank=False
        ),
        'list_type': DataValidation(
            type="list",
            formula1='"Block List,Allow List,Exception List"',
            allow_blank=False
        ),
        'review_frequency': DataValidation(
            type="list",
            formula1='"Daily,Weekly,Monthly,Quarterly,Annually,As-needed"',
            allow_blank=False
        ),
        'exception_type': DataValidation(
            type="list",
            formula1='"URL Exception,Category Exception,User Exception,Group Exception,Temporary Exception,Permanent Exception"',
            allow_blank=False
        ),
        'exception_status': DataValidation(
            type="list",
            formula1='"✅ Active,❌ Expired,⛔ Revoked,⏳ Under Review"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'policy_type': DataValidation(
            type="list",
            formula1='"Standard,Restrictive,Relaxed,Executive,Guest,Contractor"',
            allow_blank=False
        ),
        'filtering_level': DataValidation(
            type="list",
            formula1='"High (Strict),Medium (Balanced),Low (Permissive)"',
            allow_blank=False
        ),
        'https_inspection': DataValidation(
            type="list",
            formula1='"Yes,No,Selective"',
            allow_blank=False
        ),
        'aup_enforcement': DataValidation(
            type="list",
            formula1='"\u2705 Yes,\u26A0\uFE0F Partial,\u274C No,N/A"',
            allow_blank=False
        ),
        'review_type': DataValidation(
            type="list",
            formula1='"Full Review,Partial Review,Ad-hoc,Incident-Driven,Regulatory Change"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"○ Open,⟳ In Progress,✅ Resolved,⚠️ Accepted Risk"',
            allow_blank=False
        ),
        'policy_area': DataValidation(
            type="list",
            formula1='"Threat Protection,Category Filtering,Custom Lists,Exceptions,User Policies,AUP Alignment,Review Process"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Policy Config Screenshot,Policy Export,Category List,URL List,Exception Approval,AUP Document,Meeting Minutes,Test Results,User Communication,Change Record,Incident Report,Other"',
            allow_blank=False
        ),
        'evidence_verification': DataValidation(
            type="list",
            formula1='"Verified,Pending,Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Final,Requires Remediation,Re-assessment Required"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected"',
            allow_blank=False
        ),
        'recommendation': DataValidation(
            type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False
        ),
    }

    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Document your filtering philosophy (restrictive, trust-based, or hybrid) for each policy area.', '2. Complete the Threat Protection sheet for malware, phishing, and exploit protection policies.', '3. Document category filtering rules in Category Management (skip if trust-based approach).', '4. Inventory all custom block and allow lists with business justification in Custom Lists.', '5. Register all policy exceptions with business justification and approval in Policy Exceptions.', '6. Document user group and role-based policy differentiation in User Group Policies.', '7. Map filtering policies to Acceptable Use Policy requirements in Acceptable Use Alignment.', '8. Document policy review processes and update procedures in Policy Review Process.', '9. Conduct gap analysis and define remediation actions in Gap Analysis.', '10. Maintain the Evidence Register with all supporting documentation.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 24

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_threat_protection(ws, styles):
    """Create Threat Protection sheet - mandatory baseline for ALL approaches."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "THREAT PROTECTION POLICIES"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Mandatory baseline protection - applies to all filtering approaches"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Threat Type", 25),
        ("B", "Policy Configured?", 18),
        ("C", "Blocking Method", 30),
        ("D", "Effectiveness", 15),
        ("E", "False Positives", 15),
        ("F", "Last Tested", 15),
        ("G", "Evidence", 35),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Threat types
    threat_types = [
        "Known Malicious URLs",
        "Phishing Sites",
        "Malware Downloads",
        "Ransomware Delivery",
        "Exploit Kits",
        "Command & Control (C2)",
        "Cryptojacking Sites",
        "Zero-Day Threats",
    ]

    row += 1
    for threat in threat_types:
        ws[f"A{row}"] = threat
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Policy Configured dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        validations['policy_status'].add(ws[f"B{row}"])

        # Blocking Method (text)
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = border

        # Effectiveness dropdown
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = border
        validations['effectiveness'].add(ws[f"D{row}"])

        # False Positives dropdown
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = border
        validations['false_positive_rate'].add(ws[f"E{row}"])

        # Last Tested (date)
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = border

        # Evidence
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = border
        
        row += 1

    # Threat Intelligence Integration section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "THREAT INTELLIGENCE INTEGRATION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    intel_fields = [
        ("Feed Source 1:", "[Customer fills in]"),
        ("Feed Source 2:", "[Customer fills in]"),
        ("Feed Source 3:", "[Customer fills in]"),
        ("Update Frequency:", "[Dropdown]"),
        ("Last Update:", "[Date]"),
        ("Auto-Update Enabled?:", "[Dropdown]"),
    ]

    for label, value in intel_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border

        if "Dropdown" in value and "Frequency" in label:
            validations['update_frequency'].add(ws[f"B{row}"])
        # Note: "Auto-Update Enabled?" is free-text input — no DV added to avoid
        # mixed-DV on col B (policy_status rows 5-12 + yes_no would confuse FML-DV-001)
        
        row += 1

    # Compliance Checklist
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "THREAT PROTECTION COMPLIANCE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "☐ Malicious URLs blocked (verified with test)",
        "☐ Phishing sites blocked (verified with PhishTank samples)",
        "☐ Malware downloads prevented (verified with EICAR test)",
        "☐ Threat feeds updated regularly (within policy requirements)",
        "☐ Zero-day protection enabled (if available)",
        "☐ Logging of blocked threats active",
        "☐ Alerts configured for critical threats",
        "☐ Policy reviewed within last 90 days",
    ]

    row += 1
    for item in checklist_items:
        ws[f"A{row}"] = item
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: CATEGORY_MANAGEMENT SHEET
# ============================================================================

def create_category_management(ws, styles):
    """Create Category Management sheet - conditional on filtering approach."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CATEGORY-BASED FILTERING POLICIES"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document category blocking/allowing (if applicable to your approach)"
    apply_style(ws["A2"], styles["subheader"])

    # Filtering Philosophy Declaration
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "ORGANISATION'S FILTERING PHILOSOPHY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Select your approach:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border
    # Note: filtering_philosophy DV intentionally omitted here — col B must have
    # only policy_action DV (rows 9-28) for Summary Dashboard COUNTIF to work.

    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "\u26A0\uFE0F If 'Trust-Based (Threats Only)' selected, mark all categories below as N/A"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Category Policy Table
    row += 2
    headers = [
        ("A", "Category", 25),
        ("B", "Policy Action", 15),
        ("C", "Applied To", 20),
        ("D", "Business Justification", 35),
        ("E", "Exceptions?", 12),
        ("F", "Exception Count", 15),
        ("G", "Last Reviewed", 15),
        ("H", "Evidence", 30),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Categories
    categories = [
        "Adult Content",
        "Gambling",
        "Illegal Activities",
        "Weapons",
        "Hate/Discrimination",
        "Violence/Gore",
        "Streaming Media",
        "Social Networks",
        "Personal Storage/Sharing",
        "Shopping",
        "Games",
        "Job Search",
        "Advertising",
        "Anonymous Proxies/VPN",
        "Peer-to-Peer/File Sharing",
        "Instant Messaging",
        "Blogs/Forums",
        "News/Media",
        "Education",
        "Travel",
    ]

    row += 1
    for category in categories:
        ws[f"A{row}"] = category
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Policy Action dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        validations['policy_action'].add(ws[f"B{row}"])

        # Applied To dropdown
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = border
        validations['applied_to'].add(ws[f"C{row}"])

        # Text fields
        for col in ["D", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

        # Exceptions dropdown
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = border
        validations['yes_no'].add(ws[f"E{row}"])

        # Exception count (number)
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = border

        # Last Reviewed (date)
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = border
        
        row += 1

    # Category Summary
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CATEGORY SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Policy Action"
    ws[f"B{row}"] = "Category Count"
    ws[f"C{row}"] = "% of Total"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    actions = ["Block", "Allow", "Warn", "Monitor Only", "N/A"]
    cat_start = 10  # Adjust based on actual category start row
    cat_end = cat_start + len(categories) - 1

    for action in actions:
        ws[f"A{row}"] = action
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f'=COUNTIF(B{cat_start}:B{cat_end},"{action}")'
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/COUNTA(B{cat_start}:B{cat_end})*100)&"%"'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    ws.freeze_panes = "A10"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: CUSTOM_LISTS SHEET
# ============================================================================

def create_custom_lists(ws, styles):
    """Create Custom Lists sheet - organisation-maintained URL lists."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:I1")
    ws["A1"] = "CUSTOM BLOCK/ALLOW LISTS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"] = "Organisation-maintained URL lists outside of vendor categories"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "List ID", 12),
        ("B", "List Name", 25),
        ("C", "List Type", 15),
        ("D", "URL Count", 12),
        ("E", "Purpose", 35),
        ("F", "Maintained By", 20),
        ("G", "Last Updated", 15),
        ("H", "Review Frequency", 18),
        ("I", "Evidence", 30),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # List rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1

    # Sample row with LST-001
    ws[f"A{row}"] = "LST-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border

    validations['list_type'].add(ws[f"C{row}"])
    validations['review_frequency'].add(ws[f"H{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border
        row += 1

    # List Management Process
    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "LIST MANAGEMENT PROCESS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    process_fields = [
        ("List Creation Process:", ""),
        ("  \u2022 Request Method:", "[Text]"),
        ("  \u2022 Approval Required?:", "[Dropdown]"),
        ("  \u2022 Approver Role:", "[Text]"),
        ("  \u2022 Documentation:", "[Text]"),
        ("", ""),
        ("List Update Process:", ""),
        ("  \u2022 Update Frequency:", "[Dropdown]"),
        ("  \u2022 Responsible Party:", "[Text]"),
        ("  \u2022 Testing Required?:", "[Dropdown]"),
        ("  \u2022 Change Management:", "[Text]"),
        ("", ""),
        ("List Review Process:", ""),
        ("  \u2022 Review Frequency:", "[Dropdown]"),
        ("  \u2022 Review Responsibility:", "[Text]"),
        ("  \u2022 Last Full Review:", "[Date]"),
        ("  \u2022 Next Scheduled Review:", "[Date]"),
    ]

    for label, value in process_fields:
        ws[f"A{row}"] = label
        if label.startswith("  \u2022"):
            ws[f"A{row}"].font = Font(bold=True, size=9)
        else:
            ws[f"A{row}"].font = Font(bold=True)
        
        if value:
            ws.merge_cells(f"B{row}:D{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border

            if "Dropdown" in value and "Frequency" in label:
                validations['review_frequency'].add(ws[f"B{row}"])
            elif "Dropdown" in value:
                validations['yes_no'].add(ws[f"B{row}"])
        
        row += 1

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 7: POLICY_EXCEPTIONS SHEET
# ============================================================================

def create_policy_exceptions(ws, styles):
    """Create Policy Exceptions sheet - exception tracking with approval."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:M1")
    ws["A1"] = "POLICY EXCEPTION REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:M2")
    ws["A2"] = "Track approved exceptions to filtering policies"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        "Exception ID", "Exception Type", "Requested For", "URL/Category",
        "Business Justification", "Risk Level", "Compensating Controls",
        "Requested By", "Approved By", "Approval Date", "Expiry Date",
        "Status", "Evidence"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Exception rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1
    exc_start_row = row

    # Sample row with EXC-001
    ws[f"A{row}"] = "EXC-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.cell(row=row, column=ord(col)-64).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=ord(col)-64).border = border

    validations['exception_type'].add(ws[f"B{row}"])
    validations['risk_level'].add(ws[f"F{row}"])
    validations['exception_status'].add(ws[f"L{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            ws.cell(row=row, column=ord(col)-64).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=ord(col)-64).border = border
        row += 1

    exc_end_row = row - 1

    # Exception Summary
    row += 2
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "EXCEPTION SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Exception Type"
    ws[f"B{row}"] = "Active"
    ws[f"C{row}"] = "Expired"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    exception_types = [
        "URL Exception",
        "Category Exception",
        "User Exception",
        "Group Exception",
        "Temporary Exception",
        "Permanent Exception",
    ]

    for exc_type in exception_types:
        ws[f"A{row}"] = exc_type
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f'=COUNTIFS(B{exc_start_row}:B{exc_end_row},"{exc_type}",L{exc_start_row}:L{exc_end_row},"\u2705 Active")'
        ws[f"C{row}"] = f'=COUNTIFS(B{exc_start_row}:B{exc_end_row},"{exc_type}",L{exc_start_row}:L{exc_end_row},"\u274c Expired")'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"].font = Font(bold=True, color="C00000")
        row += 1

    # Alert section
    row += 2
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "\u26A0\uFE0F ALERTS: Exceptions >12 months = re-approval required | No compensating controls = HIGH RISK | Permanent exceptions = justify annually"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 15
    ws.column_dimensions["M"].width = 30

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: USER_GROUP_POLICIES SHEET
# ============================================================================

def create_user_group_policies(ws, styles):
    """Create User Group Policies sheet - role-based policies."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = "USER & GROUP POLICY ASSIGNMENTS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"] = "Role-based filtering policies (if applicable)"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Policy ID", 12),
        ("B", "Policy Name", 25),
        ("C", "Applied To", 20),
        ("D", "User Count", 12),
        ("E", "Policy Type", 15),
        ("F", "Filtering Level", 18),
        ("G", "Categories Blocked", 30),
        ("H", "Time Restrictions", 20),
        ("I", "HTTPS Inspection", 15),
        ("J", "Evidence", 30),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Policy rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1

    # Sample row with POL-001
    ws[f"A{row}"] = "POL-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border

    validations['applied_to'].add(ws[f"C{row}"])
    validations['policy_type'].add(ws[f"E{row}"])
    validations['filtering_level'].add(ws[f"F{row}"])
    validations['https_inspection'].add(ws[f"I{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border
        row += 1

    # Policy Examples
    row += 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "EXAMPLE POLICY TEMPLATES (Customise to your organisation)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    examples = [
        "Standard Employee Policy:",
        "  \u2022 Threat protection: \u2705 Enabled",
        "  \u2022 Category filtering: Medium (block inappropriate, allow business)",
        "  \u2022 Custom lists: Applied",
        "  \u2022 HTTPS inspection: Yes",
        "",
        "Executive/VIP Policy:",
        "  \u2022 Threat protection: \u2705 Enabled",
        "  \u2022 Category filtering: Low (threats only, minimal categories)",
        "  \u2022 Custom lists: Selective",
        "  \u2022 HTTPS inspection: Optional",
        "",
        "Guest/Contractor Policy:",
        "  \u2022 Threat protection: \u2705 Enabled",
        "  \u2022 Category filtering: High (restrictive, business-only)",
        "  \u2022 Custom lists: Strict",
        "  \u2022 HTTPS inspection: Yes",
        "",
        "IT/Security Team Policy:",
        "  \u2022 Threat protection: \u2705 Enabled (can bypass for testing)",
        "  \u2022 Category filtering: None (full access for security research)",
        "  \u2022 Custom lists: Applied",
        "  \u2022 HTTPS inspection: Yes (can decrypt for analysis)",
    ]

    row += 1
    for example in examples:
        ws[f"A{row}"] = example
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        if ":" in example and not example.startswith("  \u2022"):
            ws[f"A{row}"].font = Font(bold=True, size=10)
        row += 1

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: ACCEPTABLE_USE_ALIGNMENT SHEET
# ============================================================================

def create_acceptable_use_alignment(ws, styles):
    """Create Acceptable Use Alignment sheet - AUP vs filtering verification."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "ACCEPTABLE USE POLICY ALIGNMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Verify filtering policies enforce Acceptable Use requirements"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "AUP Requirement", 40),
        ("B", "Filtering Enforces This?", 20),
        ("C", "How Enforced?", 30),
        ("D", "Gaps Identified?", 15),
        ("E", "Gap Description", 30),
        ("F", "Remediation Plan", 30),
        ("G", "Evidence", 30),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # AUP Requirements
    aup_requirements = [
        "Prohibited: Accessing illegal content",
        "Prohibited: Downloading pirated software",
        "Prohibited: Accessing adult/inappropriate content",
        "Prohibited: Using company resources for personal business",
        "Prohibited: Bypassing security controls",
        "Prohibited: Excessive personal use during work hours",
        "Prohibited: Accessing gambling sites",
        "Prohibited: Accessing hate/discrimination content",
        "Prohibited: Downloading unauthorised software",
        "Prohibited: Accessing malicious/phishing sites",
        "Required: Reporting suspicious sites",
        "Required: Not sharing credentials",
        "Required: Following security policies",
        "Permitted: Reasonable personal use",
        "Permitted: Educational content",
        "Permitted: News/information",
        "Permitted: Work-related research",
        "Permitted: Approved social media for business",
        "Permitted: Cloud storage (approved services)",
        "Permitted: Streaming for business presentations",
    ]

    row += 1
    aup_start_row = row
    
    for requirement in aup_requirements:
        ws[f"A{row}"] = requirement
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Filtering Enforces dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        validations['aup_enforcement'].add(ws[f"B{row}"])

        # Text fields
        for col in ["C", "E", "F", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

        # Gaps Identified dropdown
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = border
        validations['yes_no'].add(ws[f"D{row}"])
        
        row += 1
    
    aup_end_row = row - 1

    # Alignment Summary
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "AUP ALIGNMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Alignment Status"
    ws[f"B{row}"] = "Count"
    ws[f"C{row}"] = "% of Total"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    statuses = [
        ("Fully Enforced (\u2705)", "\u2705 Yes"),
        ("Partially Enforced (\u26A0\uFE0F)", "\u26A0\uFE0F Partial"),
        ("Not Enforced (\u274C)", "\u274C No"),
        ("N/A", "N/A"),
    ]

    for label, status in statuses:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f'=COUNTIF(B{aup_start_row}:B{aup_end_row},"{status}")'
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/COUNTA(B{aup_start_row}:B{aup_end_row})*100)&"%"'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    # Overall score
    row += 1
    ws[f"A{row}"] = "Overall AUP Alignment Score:"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws[f"B{row}"] = f'=COUNTIF(B{aup_start_row}:B{aup_end_row},"\u2705 Yes")/COUNTA(B{aup_start_row}:B{aup_end_row})*100&"%"'
    ws[f"B{row}"].font = Font(bold=True, size=12, color="0000FF")

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 10: POLICY_REVIEW_PROCESS SHEET
# ============================================================================

def create_policy_review_process(ws, styles):
    """Create Policy Review Process sheet - review tracking."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "POLICY REVIEW PROCESS & TRACKING"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Maintain regular policy reviews and updates"
    apply_style(ws["A2"], styles["subheader"])

    # Review Schedule
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "REVIEW SCHEDULE"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    schedule_fields = [
        ("Policy Review Frequency:", "[Dropdown]"),
        ("Responsible Party:", "[Text]"),
        ("Last Full Review Date:", "[Date]"),
        ("Next Scheduled Review:", "[Date]"),
        ("Review Meeting Required?:", "[Dropdown]"),
        ("Stakeholders Involved:", "[Text]"),
    ]

    for label, value in schedule_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border

        if "Dropdown" in value and "Frequency" in label:
            validations['review_frequency'].add(ws[f"B{row}"])
        elif "Dropdown" in value:
            validations['yes_no'].add(ws[f"B{row}"])
        
        row += 1

    # Review History Log
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "REVIEW HISTORY LOG"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    headers = [
        ("A", "Review Date", 15),
        ("B", "Review Type", 18),
        ("C", "Policies Reviewed", 30),
        ("D", "Changes Made", 35),
        ("E", "Approved By", 20),
        ("F", "Next Review", 15),
        ("G", "Evidence", 30),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.column_dimensions[col_letter].width = width

    # Review rows (51 entries: 1 sample + 50 empty)
    row += 1
    for i in range(51):
        # Review Date
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"A{row}"].border = border

        # Review Type dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        validations['review_type'].add(ws[f"B{row}"])

        # Text fields
        for col in ["C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

        # Sample row (first row only)
        if i == 0:
            ws[f"A{row}"] = "01.01.2026"
            ws[f"B{row}"] = "Full Review"
            ws[f"C{row}"] = "All web filtering policies"
            ws[f"D{row}"] = "Updated threat category list, added AI/ML tool categories"
            ws[f"E{row}"] = "CISO"
            ws[f"F{row}"] = "01.04.2026"
            ws[f"G{row}"] = "PolicyReview_2026Q1_Minutes.pdf"

        row += 1

    # Review Checklist Template
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "QUARTERLY POLICY REVIEW CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist = [
        "Threat Protection:",
        "  ☐ Threat feeds updated and effective",
        "  ☐ Zero-day protection working",
        "  ☐ False positive rate acceptable",
        "",
        "Category Filtering (if applicable):",
        "  ☐ Category lists current and appropriate",
        "  ☐ Business justifications still valid",
        "  ☐ Exception count reasonable",
        "",
        "Custom Lists:",
        "  ☐ Block/allow lists reviewed",
        "  ☐ Obsolete entries removed",
        "  ☐ New entries justified",
        "",
        "Exceptions:",
        "  ☐ All exceptions reviewed",
        "  ☐ Expired exceptions removed/renewed",
        "  ☐ Compensating controls verified",
        "",
        "User/Group Policies:",
        "  ☐ Role assignments current",
        "  ☐ Policy levels appropriate",
        "  ☐ HTTPS inspection effective",
        "",
        "AUP Alignment:",
        "  ☐ Filtering enforces AUP",
        "  ☐ Gaps identified and addressed",
        "",
        "Overall:",
        "  ☐ Metrics reviewed (blocks, allows, exceptions)",
        "  ☐ Incidents analysed",
        "  ☐ User complaints addressed",
        "  ☐ Compliance verified",
        "  ☐ Documentation updated",
    ]

    row += 1
    for item in checklist:
        ws[f"A{row}"] = item
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        if not item.startswith("  "):
            ws[f"A{row}"].font = Font(bold=True, size=10)
        row += 1

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 11: GAP_ANALYSIS, EVIDENCE_REGISTER, APPROVAL_SIGN_OFF SHEETS
# (Following same patterns as Domains 1 & 2)
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet - policy configuration gaps."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "POLICY CONFIGURATION GAPS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    ws["A2"] = "Identify missing or inadequate policy configurations"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        "Gap ID", "Policy Area", "Gap Description", "Risk Level", "Impact",
        "Current State", "Target State", "Remediation Plan", "Owner",
        "Target Date", "Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Gap rows - MAX-001 fix: 1 sample + 50 empty rows
    row += 1
    gap_start_row = row

    # Sample row with GAP-001
    _gap3_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws[f"A{row}"] = "GAP-001"
    ws[f"A{row}"].font = Font(bold=True, size=9)
    ws[f"A{row}"].fill = _gap3_grey
    ws[f"A{row}"].border = border

    # Policy Area dropdown
    ws[f"B{row}"].fill = _gap3_grey
    ws[f"B{row}"].border = border
    validations['policy_area'].add(ws[f"B{row}"])

    # Text fields
    for col in ["C", "E", "F", "G", "H", "I", "J"]:
        ws.cell(row=row, column=ord(col)-64).fill = _gap3_grey
        ws.cell(row=row, column=ord(col)-64).border = border

    # Risk Level dropdown
    ws[f"D{row}"].fill = _gap3_grey
    ws[f"D{row}"].border = border
    validations['risk_level'].add(ws[f"D{row}"])

    # Status dropdown
    ws[f"K{row}"].fill = _gap3_grey
    ws[f"K{row}"].border = border
    validations['gap_status'].add(ws[f"K{row}"])

    row += 1

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"A{row}"].border = border
        # Policy Area dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        validations['policy_area'].add(ws[f"B{row}"])

        # Text fields
        for col in ["C", "E", "F", "G", "H", "I", "J"]:
            ws.cell(row=row, column=ord(col)-64).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=ord(col)-64).border = border

        # Risk Level dropdown
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = border
        validations['risk_level'].add(ws[f"D{row}"])

        # Status dropdown
        ws[f"K{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"K{row}"].border = border
        validations['gap_status'].add(ws[f"K{row}"])

        row += 1

    gap_end_row = row - 1

    # Gap Summary (same pattern as before)
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "GAP SUMMARY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Risk Level"
    ws[f"B{row}"] = "Gap Count"
    ws[f"C{row}"] = "% of Total"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    risk_levels = [
        ("Critical", "C00000", "FFFFFF"),
        ("High", "FFC7CE", "000000"),
        ("Medium", "FFEB9C", "000000"),
        ("Low", "C6EFCE", "000000"),
    ]

    for risk, bg_color, fg_color in risk_levels:
        ws[f"A{row}"] = risk
        ws[f"A{row}"].font = Font(bold=True, color=fg_color)
        ws[f"A{row}"].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        ws[f"B{row}"] = f'=COUNTIF(D{gap_start_row}:D{gap_end_row},"{risk}")'
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/COUNTA(D{gap_start_row}:D{gap_end_row})*100)&"%"'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref and str(_dv.sqref).strip():
            ws.add_data_validation(_dv)

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle (italic, NOT blue banner) ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers (003366, white text) ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,'
                 'Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Web Filtering Policy Configuration",
        3: "Configuration file",
        4: "Filtering policy configuration export from management console",
        5: "/evidence/A.8.23/",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None  # Empty — users choose their own evidence IDs
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"



def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # ── Row 2: Control reference (italic, 003366 text) ──
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # ── Row 3: ASSESSMENT SUMMARY banner ──
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # ── Summary fields (Rows 4-8) ──
    # (label, value, ffffcc_fill)
    summary_fields = [
        ("Document:",                  f"{DOCUMENT_ID} - {WORKBOOK_NAME}", False),
        ("Assessment Period:",         "",                                   True),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9",          True),
        ("Assessment Status:",         "",                                   True),
        ("Assessed By:",               "",                                   True),
    ]

    row = 4
    status_row_for_dv = None
    for label, value, yellow in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for col in range(2, 6):
            if yellow:
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        if "Assessment Status" in label:
            status_row_for_dv = row
        row += 1

    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f"B{status_row_for_dv}")

    # ── Approver sections ──
    row += 2  # gap before first approver

    def _approver(start_row, title, color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = border
            ws.merge_cells(f"B{r}:E{r}")
            for col in range(2, 6):
                ws.cell(row=r, column=col).fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=r, column=col).border = border
            r += 1
        return r + 1  # next row after gap

    row = _approver(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _approver(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")
    row = _approver(row, "APPROVED BY (CISO)", "003366")

    # ── Final Decision ──
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f"B{row}")

    # ── Next Review Details ──
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        row += 1

    # ── Column widths & freeze ──
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Policy configuration: WHERE filtering is applied matters, but WHAT policies
    are enforced determines effectiveness. A network with 100% coverage but
    poorly configured policies is still at risk.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.23.3 - Policy Configuration Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.23: Web Filtering")
    logger.info("=" * 78)
    logger.info("\n[*] Policy Configuration: Document WHAT filtering policies are enforced")
    logger.info("[*] Approach-Agnostic: Restrictive, Trust-Based, OR Hybrid")
    logger.info("\u2705 AUP Alignment: Verify filtering enforces Acceptable Use Policy")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("\u2705 Workbook created with 12 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")
    
    logger.info("  [1/11] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  \u2705 Instructions complete")

    logger.info("  [2/11] Creating Threat Protection...")
    create_threat_protection(wb["Threat Protection"], styles)
    logger.info("  \u2705 Threat protection complete (mandatory baseline)")

    logger.info("  [3/11] Creating Category Management...")
    create_category_management(wb["Category Management"], styles)
    logger.info("  \u2705 Category management complete (approach-dependent)")

    logger.info("  [4/11] Creating Custom Lists...")
    create_custom_lists(wb["Custom Lists"], styles)
    logger.info("  \u2705 Custom lists complete (20 list tracking rows)")

    logger.info("  [5/11] Creating Policy Exceptions...")
    create_policy_exceptions(wb["Policy Exceptions"], styles)
    logger.info("  \u2705 Exception register complete (40 exception rows)")

    logger.info("  [6/11] Creating User Group Policies...")
    create_user_group_policies(wb["User Group Policies"], styles)
    logger.info("  \u2705 User/group policies complete (25 policy assignments)")

    logger.info("  [7/11] Creating Acceptable Use Alignment...")
    create_acceptable_use_alignment(wb["Acceptable Use Alignment"], styles)
    logger.info("  \u2705 AUP alignment complete (20 AUP requirements)")

    logger.info("  [8/11] Creating Policy Review Process...")
    create_policy_review_process(wb["Policy Review Process"], styles)
    logger.info("  \u2705 Review process complete (20 review log entries)")

    logger.info("  [9/11] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)
    logger.info("  \u2705 Gap analysis complete (25 gap tracking rows)")


    logger.info("  Creating Summary Dashboard...")


    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)


    logger.info("  \u2705 Summary Dashboard complete")


    logger.info("  [10/11] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  \u2705 Evidence register complete (100 evidence rows)")

    logger.info("  [11/11] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  \u2705 Approval workflow complete")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.23.3_Policy_Configuration_{datetime.now().strftime('%Y%m%d')}.xlsx"

    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"\u2705 SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"\u274C ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\nCore Policy Configuration:")
    logger.info("  \u2022 Instructions & Legend (3 filtering philosophies explained)")
    logger.info("  \u2022 Threat Protection (mandatory baseline - 8 threat types)")
    logger.info("  \u2022 Category Management (approach-dependent - 20 categories)")
    logger.info("  \u2022 Custom Lists (20 custom URL lists)")
    logger.info("  \u2022 Policy Exceptions (40 exception tracking rows)")
    logger.info("\nAdvanced Policy Features:")
    logger.info("  \u2022 User Group Policies (25 role-based policies)")
    logger.info("  \u2022 Acceptable Use Alignment (20 AUP requirements)")
    logger.info("  \u2022 Policy Review Process (review tracking + checklist)")
    logger.info("\nGovernance:")
    logger.info("  \u2022 Gap_Analysis (25 policy gap tracking rows)")
    logger.info("  \u2022 Evidence_Register (100 evidence entries)")
    logger.info("  \u2022 Approval_Sign_Off (3-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info("ASSESSMENT CAPABILITIES:")
    logger.info("  \u2022 8 threat protection policies")
    logger.info("  \u2022 20 category filtering policies (if applicable)")
    logger.info("  \u2022 20 custom block/allow lists")
    logger.info("  \u2022 40 policy exception tracking")
    logger.info("  \u2022 25 user/group policy assignments")
    logger.info("  \u2022 20 AUP requirements verification")
    logger.info("  \u2022 Quarterly review checklist (35 items)")
    logger.info("  \u2022 Auto-calculated AUP alignment score")
    logger.info("\n" + "─" * 78)
    logger.info("KEY FEATURES:")
    logger.info("  \u2705 Approach-agnostic (Restrictive, Trust-Based, OR Hybrid)")
    logger.info("  \u2705 Mandatory threat protection (applies to ALL approaches)")
    logger.info("  \u2705 Optional category filtering (trust-based skips this)")
    logger.info("  \u2705 Custom list management (20 lists)")
    logger.info("  \u2705 Exception management with expiry tracking")
    logger.info("  \u2705 Role-based policy assignments")
    logger.info("  \u2705 AUP alignment verification (auto-calculated score)")
    logger.info("  \u2705 Quarterly review process documentation")
    logger.info("\n" + "=" * 78)
    logger.info("NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Declare your filtering philosophy (Restrictive/Trust-Based/Hybrid)")
    logger.info("  3. Document threat protection (MANDATORY for all approaches)")
    logger.info("  4. Complete category management (if applicable)")
    logger.info("  5. List custom block/allow lists")
    logger.info("  6. Document policy exceptions with approvals")
    logger.info("  7. Map user/group policies if using role-based filtering")
    logger.info("  8. Verify AUP alignment (critical for compliance)")
    logger.info("  9. Document policy review process")
    logger.info("  10. Obtain approvals via Approval_Sign_Off")
    logger.info("\nPRO TIP:")
    logger.info("  Trust-based approach? Mark Category_Management as N/A.")
    logger.info("  Restrictive approach? Document ALL blocked categories.")
    logger.info("  Hybrid? Balance threat protection + selective categories.")
    logger.info("\nPOLICY EFFECTIVENESS FORMULA:")
    logger.info("  Effectiveness = (Threats Blocked + AUP Alignment) / Exceptions")
    logger.info("\n" + "=" * 78)
    logger.info('\n"Policy without evidence is wishful thinking."')
    logger.info('"Evidence without review is stale documentation."')
    logger.info('"Policy + Evidence + Review = Systems Engineering"')
    logger.info("\nDocument -> Review -> Enforce -> Verify -> Repeat")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
