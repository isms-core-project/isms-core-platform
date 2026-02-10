#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-IMP-A.8.23.4 - Monitoring & Response Assessment
Comprehensive Data Population Script for CISO Presentation

Populates ALL sheets with production-quality demo data:
- Log Collection (sources and configuration)
- Alert Configuration (thresholds and notifications)
- Monitoring Dashboard (metrics and KPIs)
- Incident Response (procedures and playbooks)
- Blocked Events Analysis (trends and patterns)
- False Positive Management (tuning records)
- Reporting Schedule (stakeholder reports)
- Gap Analysis
- Evidence Register (25+ documents)
- Approval Sign-Off

Usage:
    python3 populate_a823_4_monitoring_response.py ISMS-IMP-A.8.23.4.xlsx

Generated: 2026-02-01
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def populate_log_collection(wb):
    """Populate Log Collection configuration."""
    ws = wb["Log_Collection"]

    data = [
        # Log Source, Log Type, Collection Method, Destination, Retention, Format, Volume/Day, Status
        ["Zscaler ZIA", "Web Traffic Logs", "NSS (Nanolog Streaming Service)", "Splunk Cloud", "90 days (hot), 1 year (cold)", "JSON/CEF", "~50 GB", "Active"],
        ["Zscaler ZIA", "Threat Logs", "NSS", "Splunk Cloud + SOAR", "1 year", "JSON/CEF", "~2 GB", "Active"],
        ["Zscaler ZIA", "DNS Logs", "NSS", "Splunk Cloud", "90 days", "JSON/CEF", "~10 GB", "Active"],
        ["Zscaler ZIA", "Firewall Logs", "NSS", "Splunk Cloud", "90 days", "JSON/CEF", "~5 GB", "Active"],
        ["Zscaler ZIA", "DLP Logs", "NSS", "Splunk Cloud + DLP Console", "1 year", "JSON/CEF", "~500 MB", "Active"],
        ["Zscaler ZPA", "Access Logs", "LSS (Log Streaming Service)", "Splunk Cloud", "90 days", "JSON/CEF", "~3 GB", "Active"],
        ["Zscaler ZCC", "Client Logs", "ZCC telemetry", "Zscaler Console", "30 days", "Proprietary", "~1 GB", "Active"],
        ["Prisma Access", "Traffic Logs", "Syslog forwarding", "Splunk Cloud", "90 days", "CEF", "~5 GB", "Active"],
        ["Cisco Umbrella", "DNS Query Logs", "S3 export", "Splunk Cloud", "30 days", "CSV", "~2 GB", "Active - Legacy"],
        ["Microsoft Defender", "Endpoint Alerts", "Microsoft 365 Defender API", "Splunk Cloud + Sentinel", "90 days", "JSON", "~500 MB", "Active"],
        ["Microsoft Entra ID (formerly Azure AD)", "Sign-in Logs", "Diagnostic Settings", "Splunk Cloud + Sentinel", "90 days", "JSON", "~1 GB", "Active"],
        ["Splunk (Aggregated)", "Correlated Events", "Internal correlation", "Splunk ES", "1 year", "Splunk format", "~5 GB", "Active"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Log_Collection: {count} log sources configured")
    return count


def populate_alert_configuration(wb):
    """Populate Alert Configuration."""
    ws = wb["Alert_Configuration"]

    data = [
        # Alert Name, Trigger Condition, Severity, Threshold, Response Action, Notification, SLA, Status
        ["Malware Detected", "Malware file blocked", "Critical", "Any detection", "Auto-ticket, endpoint scan trigger", "SOC + Endpoint Team", "15 min response", "Active"],
        ["Ransomware Indicator", "Ransomware behavior detected", "Critical", "Any detection", "Auto-isolate endpoint, SOC alert", "SOC + CISO + Endpoint", "5 min response", "Active"],
        ["C2 Communication", "C2 callback blocked", "Critical", "Any detection", "Auto-ticket, endpoint investigation", "SOC + Threat Intel", "15 min response", "Active"],
        ["Phishing Access Blocked", "Phishing URL blocked", "High", ">5/user/day", "User notification, awareness flag", "SOC + HR (if repeated)", "30 min response", "Active"],
        ["Credential Theft Attempt", "Credential phishing blocked", "Critical", "Any detection", "Password reset trigger, SOC alert", "SOC + Identity Team", "15 min response", "Active"],
        ["Data Exfiltration Attempt", "DLP policy violation", "High", "Any PII/PCI upload blocked", "Auto-ticket, manager notification", "SOC + DLP Team + Manager", "30 min response", "Active"],
        ["Proxy Bypass Attempt", "Anonymizer/proxy site access", "High", "Any attempt", "Block + disciplinary flag", "SOC + HR", "1 hour response", "Active"],
        ["Unusual Traffic Volume", "User traffic >5x baseline", "Medium", "5x 7-day average", "Investigation ticket", "SOC", "4 hour response", "Active"],
        ["New Domain Access (High Risk)", "Access to <7-day domain flagged malicious", "High", "Any access", "Block + investigation", "SOC", "30 min response", "Active"],
        ["SSL Inspection Failure", "SSL decryption error spike", "Medium", ">100/hour", "Technical investigation", "Network Team", "2 hour response", "Active"],
        ["Policy Exception Abuse", "Exception user accessing blocked content", "Medium", ">10 blocks/day", "Exception review trigger", "IT Security + Manager", "1 day response", "Active"],
        ["Guest Network Threat", "Guest device threat detected", "Medium", "Any detection", "Guest isolation, incident log", "Network Team + Facilities", "1 hour response", "Active"],
        ["Executive Account Anomaly", "Executive unusual access pattern", "High", "Deviation from baseline", "Immediate verification", "SOC + Executive Assistant", "15 min response", "Active"],
        ["Batch Block Event", ">50 users blocked same URL", "Medium", ">50 users/hour", "False positive check", "IT Security", "1 hour response", "Active"],
        ["System Health Alert", "Zscaler service degradation", "High", "Latency >100ms or errors", "Vendor escalation", "Network Team + Zscaler TAM", "30 min response", "Active"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Alert_Configuration: {count} alerts configured")
    return count


def populate_monitoring_dashboard(wb):
    """Populate Monitoring Dashboard metrics."""
    ws = wb["Monitoring_Dashboard"]

    data = [
        # Metric Name, Current Value, Target, Trend, Time Period, Data Source, Owner, Status
        ["Total Web Requests/Day", "45.2 million", ">40M baseline", "Stable (+2%)", "Last 30 days", "Zscaler Dashboard", "IT Security", "Normal"],
        ["Threats Blocked/Day", "12,450", "N/A (lower is better)", "Decreasing (-5%)", "Last 30 days", "Zscaler Threat Dashboard", "SOC Team", "Good"],
        ["Malware Blocks/Day", "342", "N/A (lower is better)", "Stable", "Last 30 days", "Zscaler Dashboard", "SOC Team", "Normal"],
        ["Phishing Blocks/Day", "1,856", "N/A (lower is better)", "Decreasing (-8%)", "Last 30 days", "Zscaler Dashboard", "SOC Team", "Good"],
        ["C2 Blocks/Day", "23", "N/A (lower is better)", "Stable", "Last 30 days", "Zscaler Dashboard", "SOC Team", "Normal"],
        ["DLP Violations/Day", "45", "<100", "Stable", "Last 30 days", "Zscaler DLP Dashboard", "DLP Team", "Good"],
        ["Policy Blocks/Day", "8,234", "N/A (informational)", "Stable", "Last 30 days", "Zscaler Dashboard", "IT Security", "Normal"],
        ["SSL Inspection Rate", "94.2%", ">90%", "Improving (+1%)", "Last 30 days", "Zscaler Dashboard", "Network Team", "Good"],
        ["Average Latency (Zurich)", "32ms", "<50ms", "Stable", "Real-time", "Synthetic Monitoring", "Network Team", "Good"],
        ["Average Latency (Remote)", "68ms", "<100ms", "Stable", "Real-time", "Synthetic Monitoring", "Network Team", "Good"],
        ["Service Availability", "99.998%", ">99.99%", "Stable", "Last 30 days", "Zscaler Status", "Network Team", "Excellent"],
        ["False Positive Rate", "0.05%", "<0.1%", "Improving (-0.01%)", "Last 30 days", "Tuning Dashboard", "IT Security", "Good"],
        ["User Complaints/Week", "12", "<25", "Decreasing (-3)", "Last 4 weeks", "ServiceNow", "IT Security", "Good"],
        ["MTTR (Security Incidents)", "23 minutes", "<30 minutes", "Improving (-5 min)", "Last 30 days", "SOAR Platform", "SOC Team", "Good"],
        ["Coverage Rate", "94.5%", ">95%", "Improving (+2%)", "Current", "Coverage Assessment", "IT Security", "Near Target"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Monitoring_Dashboard: {count} metrics tracked")
    return count


def populate_incident_response(wb):
    """Populate Incident Response procedures."""
    ws = wb["Incident_Response"]

    data = [
        # Incident Type, Playbook ID, Detection Source, Initial Response, Escalation Path, Resolution Steps, SLA, Last Test
        ["Malware Detection", "WF-IR-001", "Zscaler threat alert", "1. Verify block 2. Check endpoint 3. Open ticket", "SOC L1 → L2 → Endpoint Team", "Endpoint scan, quarantine if needed, user notify", "15 min initial, 4h resolution", "2026-01-15"],
        ["Ransomware Alert", "WF-IR-002", "Zscaler + EDR correlation", "1. Isolate endpoint 2. Alert SOC L2 3. CISO notify", "SOC L2 → CISO → IR Team", "Full isolation, forensics, recovery from backup", "5 min initial, 24h containment", "2026-01-20"],
        ["C2 Communication", "WF-IR-003", "Zscaler C2 block", "1. Verify block 2. Identify endpoint 3. Deep scan", "SOC L1 → Threat Intel → IR Team", "Endpoint forensics, IOC extraction, threat hunt", "15 min initial, 8h resolution", "2026-01-18"],
        ["Phishing Campaign", "WF-IR-004", "Multiple users blocked", "1. Identify scope 2. Block indicators 3. User comms", "SOC L1 → IT Security → Comms", "Block all variants, user awareness blast, report to vendor", "30 min initial, 2h resolution", "2026-01-22"],
        ["Data Exfiltration", "WF-IR-005", "DLP alert", "1. Verify violation 2. Identify data 3. Manager notify", "DLP Team → IT Security → Legal", "Data assessment, user interview, policy enforcement", "30 min initial, 24h resolution", "2026-01-25"],
        ["Credential Theft", "WF-IR-006", "Credential phishing block", "1. Force password reset 2. Check account activity", "SOC → Identity Team → User", "Password reset, MFA verify, account audit", "15 min initial, 2h resolution", "2026-01-10"],
        ["Proxy Bypass", "WF-IR-007", "Anonymizer access alert", "1. Block access 2. Identify user 3. HR notify", "IT Security → HR → Manager", "Access review, disciplinary process if repeat", "1h initial, 48h resolution", "2026-01-12"],
        ["Service Degradation", "WF-IR-008", "Latency/availability alert", "1. Check Zscaler status 2. Verify scope 3. Failover if needed", "Network Team → Zscaler TAM", "Traffic reroute, vendor escalation, root cause", "30 min initial, 4h resolution", "2026-01-28"],
        ["Mass Block Event", "WF-IR-009", "Batch block alert", "1. Identify URL/category 2. Check for false positive", "IT Security → Category review", "Verify categorization, whitelist if FP, user comms", "1h initial, 4h resolution", "2026-01-08"],
        ["Executive Targeting", "WF-IR-010", "Executive anomaly alert", "1. Verify with executive 2. Check all access 3. Enhanced monitoring", "SOC L2 → CISO → Executive", "Full account audit, enhanced protection, threat hunt", "5 min initial, 24h monitoring", "2026-01-30"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Incident_Response: {count} playbooks documented")
    return count


def populate_blocked_events_analysis(wb):
    """Populate Blocked Events Analysis."""
    ws = wb["Blocked_Events_Analysis"]

    data = [
        # Category, Blocks/Month, Trend, Top Users, Top Domains, Risk Assessment, Action Taken, Notes
        ["Malware", "10,250", "Decreasing (-5%)", "Automated scans (50%)", "cdn-malware[.]ru, badfile[.]net", "High - Active threats", "Threat intel updated", "Most from email links"],
        ["Phishing", "55,680", "Stable", "All users (distributed)", "login-verify[.]com, secure-update[.]net", "High - Credential risk", "User awareness training", "Wave attacks detected"],
        ["Adult Content", "2,340", "Stable", "Guest WiFi (80%)", "Various", "Low - Policy violation", "None needed", "Expected on guest network"],
        ["Gambling", "890", "Decreasing (-10%)", "5 repeat users", "bet365[.]com, pokerstars[.]com", "Low - Policy violation", "HR notified for repeats", "Blocked per AUP"],
        ["Streaming (Over Quota)", "12,450", "Increasing (+15%)", "Marketing, remote workers", "netflix[.]com, youtube[.]com", "Low - Bandwidth", "Bandwidth policy review", "Consider quota increase"],
        ["Social Media (Call Center)", "8,900", "Stable", "Call center users", "facebook[.]com, instagram[.]com", "Low - Productivity", "None needed", "Per department policy"],
        ["Personal Cloud Storage", "3,450", "Decreasing (-20%)", "New employees", "dropbox[.]com, drive.google.com", "Medium - Data leakage", "Onboarding updated", "Training effectiveness shown"],
        ["Hacking/Proxy Avoidance", "456", "Decreasing (-15%)", "3 users (flagged)", "hide.me, nordvpn[.]com", "High - Policy bypass", "HR escalation for 3 users", "Disciplinary process"],
        ["Uncategorized (Blocked)", "1,230", "Stable", "Developers (60%)", "Various new domains", "Medium - Unknown risk", "Category review weekly", "New tools being evaluated"],
        ["C2/Botnet", "690", "Stable", "Automated (IoT/servers)", "Various C2 domains", "Critical - Active threat", "IOC sharing with threat intel", "Legacy IoT contributing"],
        ["Cryptomining", "234", "Decreasing (-25%)", "2 infected endpoints", "coinhive alternatives", "Medium - Resource abuse", "Endpoints cleaned", "Browser-based mining"],
        ["Newly Registered Domains", "4,560", "Increasing (+10%)", "All users", "Various <30-day domains", "Medium - Emerging threats", "Threshold tuning", "Business impact review needed"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Blocked_Events_Analysis: {count} categories analyzed")
    return count


def populate_false_positive_management(wb):
    """Populate False Positive Management."""
    ws = wb["False_Positive_Mgmt"]

    data = [
        # FP ID, Date Reported, URL/Domain, Reported Category, Correct Category, Reporter, Resolution, Resolution Date, Status
        ["FP-2026-001", "2026-01-05", "partner-portal.vendor.com", "Malware", "Business/Partner", "Partner Team", "Whitelisted + reported to Zscaler", "2026-01-05", "Resolved"],
        ["FP-2026-002", "2026-01-08", "newstartup.io", "Uncategorized (blocked)", "Technology/SaaS", "Developer", "Submitted for categorization", "2026-01-09", "Resolved"],
        ["FP-2026-003", "2026-01-10", "marketing-tool.com", "Adult Content", "Marketing/Analytics", "Marketing Team", "Whitelisted + vendor notified", "2026-01-10", "Resolved"],
        ["FP-2026-004", "2026-01-12", "legal-research.ch", "Gambling", "Legal/Reference", "Legal Team", "Whitelisted + reported", "2026-01-12", "Resolved"],
        ["FP-2026-005", "2026-01-15", "cloud-provider-new.com", "Phishing", "Cloud Services", "Cloud Team", "Submitted for recategorization", "2026-01-16", "Resolved"],
        ["FP-2026-006", "2026-01-18", "customer-feedback.net", "Spam", "Business Services", "Customer Success", "Whitelisted", "2026-01-18", "Resolved"],
        ["FP-2026-007", "2026-01-20", "dev-tools-2026.io", "Hacking", "Development Tools", "Developer", "Submitted for categorization", "2026-01-22", "Resolved"],
        ["FP-2026-008", "2026-01-22", "finance-news.com", "Malware", "News/Finance", "Finance Team", "Investigated - was compromised, now clean", "2026-01-23", "Resolved"],
        ["FP-2026-009", "2026-01-25", "hr-benefits-provider.com", "Phishing", "Business Services", "HR Team", "Whitelisted after verification", "2026-01-25", "Resolved"],
        ["FP-2026-010", "2026-01-28", "new-erp-module.vendor.com", "Uncategorized", "Business Applications", "IT Operations", "Pending categorization", "Pending", "Open"],
        ["FP-2026-011", "2026-01-28", "training-platform.edu", "Adult Content", "Education", "HR Team", "Under investigation", "Pending", "Open"],
        ["FP-2026-012", "2026-01-30", "api-gateway.newservice.com", "C2", "Technology/API", "Developer", "Investigating with Zscaler", "Pending", "Open"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    False_Positive_Mgmt: {count} FP records documented")
    return count


def populate_reporting_schedule(wb):
    """Populate Reporting Schedule."""
    ws = wb["Reporting_Schedule"]

    data = [
        # Report Name, Frequency, Recipients, Content Summary, Format, Delivery Method, Owner, Last Sent, Next Due
        ["Daily Threat Summary", "Daily", "SOC Team, IT Security", "Threats blocked, top categories, anomalies", "Email + Dashboard", "Automated - 06:00", "SOC Team", "2026-02-01", "2026-02-02"],
        ["Weekly Security Digest", "Weekly", "IT Security, IT Management", "Week trends, incidents, FP summary, metrics", "PDF Report", "Email - Monday 08:00", "IT Security", "2026-01-27", "2026-02-03"],
        ["Monthly Executive Summary", "Monthly", "CISO, CIO, IT Directors", "KPIs, trends, incidents, compliance status", "PowerPoint + PDF", "Email + Meeting", "CISO Office", "2026-01-05", "2026-02-05"],
        ["Monthly Compliance Report", "Monthly", "Compliance, Internal Audit", "Policy compliance, exceptions, violations", "PDF Report", "SharePoint + Email", "IT Security", "2026-01-10", "2026-02-10"],
        ["Quarterly Board Report", "Quarterly", "Board of Directors", "Executive summary, risk posture, investments", "Board Deck", "Board Portal", "CISO", "2025-12-15", "2026-03-15"],
        ["Quarterly Vendor Review", "Quarterly", "IT Security, Procurement, Zscaler TAM", "Service performance, roadmap, issues", "Meeting + Report", "Video Call", "IT Security", "2025-12-20", "2026-03-20"],
        ["User Awareness Report", "Monthly", "HR, Training Team", "Blocked attempts by user (anonymized trends)", "PDF Report", "Email", "IT Security", "2026-01-15", "2026-02-15"],
        ["DLP Incident Report", "Weekly", "DLP Team, Legal, Compliance", "Data leakage attempts, policy violations", "Secure PDF", "Encrypted Email", "DLP Team", "2026-01-28", "2026-02-04"],
        ["Exception Status Report", "Monthly", "IT Security, CISO", "Active exceptions, pending reviews, expirations", "Excel + PDF", "Email", "IT Security", "2026-01-10", "2026-02-10"],
        ["SLA Performance Report", "Monthly", "IT Security, Vendor Management", "Zscaler SLA metrics, availability, latency", "PDF Report", "Email", "Network Team", "2026-01-05", "2026-02-05"],
        ["Incident Response Summary", "Monthly", "SOC, IT Security, CISO", "Incidents handled, MTTR, lessons learned", "PDF Report", "Email + Meeting", "SOC Lead", "2026-01-08", "2026-02-08"],
        ["Coverage Assessment Update", "Quarterly", "IT Security, Network Team, CISO", "Coverage status, gaps, remediation progress", "PowerPoint", "Email + Meeting", "IT Security", "2026-01-15", "2026-04-15"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Reporting_Schedule: {count} reports scheduled")
    return count


def populate_gap_analysis(wb):
    """Populate Gap Analysis."""
    ws = wb["Gap_Analysis"]

    data = [
        # Gap ID, Category, Description, Severity, Current State, Target State, Remediation, Owner, Due Date, Status
        ["MON-GAP-001", "Log Retention", "Hot storage limited to 90 days, cold limited to 1 year", "Medium",
         "90-day hot / 1-year cold retention", "1-year hot / 7-year cold for compliance",
         "Upgrade Splunk license, implement archive tier", "IT Operations", "2026-06-30", "Planned"],
        ["MON-GAP-002", "IoT Monitoring", "IoT traffic not fully visible in SIEM", "High",
         "Limited IoT log collection", "Full IoT traffic logging and alerting",
         "Deploy Zscaler IoT solution, integrate with SIEM", "IT Security", "2026-06-30", "In Progress"],
        ["MON-GAP-003", "Automated Response", "Limited automated response actions", "Medium",
         "Manual response for most alerts", "Automated containment for critical threats",
         "Implement SOAR playbooks for top 5 scenarios", "SOC Team", "2026-04-30", "In Progress"],
        ["MON-GAP-004", "User Behavior Analytics", "No UBA/UEBA integration", "Medium",
         "Basic anomaly alerts only", "Full user behavior analytics",
         "Evaluate Zscaler Deception + UBA add-on", "IT Security", "2026-09-30", "Planned"],
        ["MON-GAP-005", "Real-time Dashboards", "Some metrics require manual refresh", "Low",
         "Mix of real-time and batch dashboards", "All critical metrics real-time",
         "Upgrade Splunk dashboards to real-time", "SOC Team", "2026-03-31", "In Progress"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Gap_Analysis: {count} gaps identified")
    return count


def populate_evidence_register(wb):
    """Populate Evidence Register with audit evidence."""
    ws = wb["Evidence_Register"]

    data = [
        # Evidence ID, Title, Type, Related Sheet, Location, Collection Date, Owner, Retention, Status
        ["EV4-001", "Zscaler NSS Configuration", "Technical Config", "Log_Collection", "SharePoint/Evidence/Zscaler/NSS", "2026-01-28", "IT Security", "3 years", "Current"],
        ["EV4-002", "Splunk Data Input Configuration", "Technical Config", "Log_Collection", "SharePoint/Evidence/Splunk/Inputs", "2026-01-25", "IT Operations", "3 years", "Current"],
        ["EV4-003", "Log Retention Policy Document", "Policy", "Log_Collection", "SharePoint/Policies/LogRetention", "2025-06-01", "Compliance", "Permanent", "Current"],
        ["EV4-004", "Zscaler Alert Rules Export", "Technical Config", "Alert_Configuration", "SharePoint/Evidence/Zscaler/Alerts", "2026-01-28", "IT Security", "3 years", "Current"],
        ["EV4-005", "Splunk Alert Configuration", "Technical Config", "Alert_Configuration", "SharePoint/Evidence/Splunk/Alerts", "2026-01-28", "SOC Team", "3 years", "Current"],
        ["EV4-006", "SOAR Playbook Documentation", "Procedure", "Alert_Configuration", "SharePoint/SOC/Playbooks", "2026-01-20", "SOC Team", "3 years", "Current"],
        ["EV4-007", "Monitoring Dashboard Screenshots", "Evidence", "Monitoring_Dashboard", "SharePoint/Evidence/Dashboards", "2026-01-28", "IT Security", "1 year", "Current"],
        ["EV4-008", "Monthly KPI Report (Jan 2026)", "Report", "Monitoring_Dashboard", "SharePoint/Reports/KPI", "2026-01-31", "IT Security", "3 years", "Current"],
        ["EV4-009", "Incident Response Playbook Library", "Procedure", "Incident_Response", "SharePoint/SOC/IRPlaybooks", "2026-01-15", "SOC Team", "3 years", "Current"],
        ["EV4-010", "IR Tabletop Exercise Report", "Test Report", "Incident_Response", "SharePoint/SOC/Exercises", "2026-01-20", "SOC Team", "3 years", "Current"],
        ["EV4-011", "Sample Incident Tickets (Sanitized)", "Evidence", "Incident_Response", "SharePoint/Evidence/Incidents", "2026-01-28", "SOC Team", "1 year", "Current"],
        ["EV4-012", "Blocked Events Trend Report", "Report", "Blocked_Events_Analysis", "SharePoint/Reports/BlockedEvents", "2026-01-28", "IT Security", "1 year", "Current"],
        ["EV4-013", "Threat Category Analysis Q4 2025", "Report", "Blocked_Events_Analysis", "SharePoint/Reports/ThreatAnalysis", "2025-12-31", "SOC Team", "3 years", "Current"],
        ["EV4-014", "False Positive Tracking Spreadsheet", "Tracking", "False_Positive_Mgmt", "SharePoint/Security/FPTracking", "2026-01-28", "IT Security", "2 years", "Current"],
        ["EV4-015", "Zscaler Categorization Submissions", "Evidence", "False_Positive_Mgmt", "SharePoint/Evidence/Categorization", "2026-01-28", "IT Security", "1 year", "Current"],
        ["EV4-016", "Weekly Security Digest (Sample)", "Report", "Reporting_Schedule", "SharePoint/Reports/WeeklyDigest", "2026-01-27", "IT Security", "1 year", "Current"],
        ["EV4-017", "Monthly Executive Report (Jan 2026)", "Report", "Reporting_Schedule", "SharePoint/Reports/Executive", "2026-01-31", "CISO Office", "5 years", "Current"],
        ["EV4-018", "Board Security Presentation Q4 2025", "Report", "Reporting_Schedule", "SharePoint/Board/Security", "2025-12-15", "CISO", "Permanent", "Current"],
        ["EV4-019", "SLA Performance Report (Jan 2026)", "Report", "Monitoring_Dashboard", "SharePoint/Reports/SLA", "2026-01-31", "Network Team", "3 years", "Current"],
        ["EV4-020", "Splunk Search Queries Library", "Technical", "All Sheets", "Git/splunk/searches", "2026-01-25", "SOC Team", "Version controlled", "Current"],
        ["EV4-021", "Alert Tuning History Log", "Change Records", "Alert_Configuration", "ServiceNow/AlertTuning", "2026-01-28", "IT Security", "3 years", "Current"],
        ["EV4-022", "MTTR Metrics Dashboard Export", "Report", "Monitoring_Dashboard", "SharePoint/Reports/MTTR", "2026-01-28", "SOC Team", "1 year", "Current"],
        ["EV4-023", "Incident Post-Mortem Reports", "Analysis", "Incident_Response", "SharePoint/SOC/PostMortems", "2026-01-25", "SOC Team", "5 years", "Current"],
        ["EV4-024", "Coverage Gap Remediation Tracker", "Tracking", "Gap_Analysis", "SharePoint/Security/GapRemediation", "2026-01-28", "IT Security", "3 years", "Current"],
        ["EV4-025", "Vendor SLA Dashboard Access", "Access Record", "Monitoring_Dashboard", "Zscaler Portal Access", "2026-01-28", "IT Security", "1 year", "Current"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Evidence_Register: {count} evidence documents")
    return count


def safe_cell_write(ws, cell_ref, value):
    """Safely write to a cell, handling merged cells."""
    try:
        cell = ws[cell_ref]
        if not isinstance(cell, MergedCell):
            cell.value = value
    except Exception as e:  # Skip merged/protected cells
        pass


def populate_approval_signoff(wb):
    """Populate Approval Sign-Off with stakeholder workflow."""
    ws = wb["Approval_Sign_Off"]

    # Assessor information
    safe_cell_write(ws, "B5", "Daniel Meier")
    safe_cell_write(ws, "B6", "SOC Analyst Lead")
    safe_cell_write(ws, "B7", "daniel.meier@company.com")
    safe_cell_write(ws, "B8", datetime.now().strftime("%d.%m.%Y"))

    # Technical reviewer
    safe_cell_write(ws, "B12", "Sandra Koch")
    safe_cell_write(ws, "B13", "Security Operations Manager")
    safe_cell_write(ws, "B14", "sandra.koch@company.com")
    safe_cell_write(ws, "B15", (datetime.now() + timedelta(days=3)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B16", "Verified - Monitoring capabilities comprehensive. SOAR integration progressing well.")

    # Security reviewer
    safe_cell_write(ws, "B20", "Anna Müller")
    safe_cell_write(ws, "B21", "Information Security Manager")
    safe_cell_write(ws, "B22", "anna.mueller@company.com")
    safe_cell_write(ws, "B23", (datetime.now() + timedelta(days=5)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B24", "Approved - Response procedures well-documented. Recommend completing automated response project.")

    # Management approval
    safe_cell_write(ws, "B28", "Dr. Klaus Fischer")
    safe_cell_write(ws, "B29", "CISO")
    safe_cell_write(ws, "B30", "klaus.fischer@company.com")
    safe_cell_write(ws, "B31", (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B32", "Approved. Excellent monitoring coverage. Prioritize IoT visibility gap.")

    # Next review
    safe_cell_write(ws, "B36", (datetime.now() + timedelta(days=90)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B37", "Security Operations Team")

    logger.info("    Approval_Sign_Off: Complete workflow populated")


def main():
    """Main function to populate the workbook."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.23.4 - Monitoring & Response Assessment")
    logger.info("Comprehensive Data Population for CISO Presentation")
    logger.info("=" * 80)

    if len(sys.argv) < 2:
        logger.error("Usage: python3 populate_a823_4_monitoring_response.py <workbook.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    logger.info(f"Loading: {filepath}")

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        logger.error(f"Failed to load workbook: {e}")
        sys.exit(1)

    logger.info(f"Sheets found: {len(wb.sheetnames)}")

    total_entries = 0

    logger.info("[1/8] Populating Log Collection...")
    total_entries += populate_log_collection(wb)

    logger.info("[2/8] Populating Alert Configuration...")
    total_entries += populate_alert_configuration(wb)

    logger.info("[3/8] Populating Monitoring Dashboard...")
    total_entries += populate_monitoring_dashboard(wb)

    logger.info("[4/8] Populating Incident Response...")
    total_entries += populate_incident_response(wb)

    logger.info("[5/8] Populating Blocked Events Analysis...")
    total_entries += populate_blocked_events_analysis(wb)

    logger.info("[6/8] Populating False Positive Management...")
    total_entries += populate_false_positive_management(wb)

    logger.info("[7/8] Populating Reporting Schedule...")
    total_entries += populate_reporting_schedule(wb)

    logger.info("[8/8] Populating Gap Analysis, Evidence & Approval...")
    total_entries += populate_gap_analysis(wb)
    evidence_count = populate_evidence_register(wb)
    populate_approval_signoff(wb)

    wb.save(filepath)
    logger.info(f"Saved: {filepath}")

    logger.info("=" * 80)
    logger.info("DATA POPULATION COMPLETE")
    logger.info("=" * 80)
    logger.info(f"Summary:")
    logger.info(f"  - Assessment Entries: {total_entries}")
    logger.info(f"  - Evidence Documents: {evidence_count}")
    logger.info(f"  - Total Data Points: {total_entries + evidence_count}")
    logger.info("CISO-Ready: Professional monitoring & response assessment")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED - PRESENTATION MODE POPULATE SCRIPT
# QA_TOOL: Claude Code
# CHANGES: Created for CISO presentation demo data
# =============================================================================
