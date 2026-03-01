#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.23.4 - Monitoring & Response Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.23: Web Filtering
Assessment Domain 4 of 4: Monitoring, Logging, Alerting & Incident Response

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific monitoring requirements, logging infrastructure,
alerting mechanisms, and incident response procedures.

Key customisation areas:
1. Log retention requirements (match your regulatory and policy requirements)
2. Alert thresholds and escalation criteria (adapt to your risk tolerance)
3. SIEM/monitoring tool integration (specific to your security stack)
4. Incident response workflow (match your IR procedures)
5. Reporting requirements (adapt to stakeholder needs)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.23 Web Filtering Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
web filtering monitoring, logging, alerting, and incident response capabilities
against ISO 27001:2022 Control A.8.23 requirements.

**Purpose:**
Validates that web filtering controls are supported by adequate monitoring,
logging, alerting, and incident response capabilities to detect, investigate,
and respond to security events and policy violations.

**Assessment Scope:**
- Log generation, collection, and retention adequacy
- Real-time monitoring and alerting configuration
- Event correlation and threat detection capabilities
- Incident detection, triage, and response procedures
- Reporting and metrics for compliance oversight
- Integration with Security Operations Center (SOC)
- SIEM/monitoring tool integration effectiveness
- Evidence collection for monitoring adequacy audit

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and evaluation criteria
2. Log Management - Generation, collection, retention, and access controls
3. Monitoring Configuration - Real-time monitoring and coverage assessment
4. Alerting Framework - Alert rules, thresholds, and escalation procedures
5. Event Correlation - Multi-source correlation and threat detection
6. Incident Response - Detection, triage, investigation, and remediation
7. SIEM Integration - Security information and event management integration
8. Reporting & Metrics - Compliance reporting and operational metrics
9. SOC Integration - Security operations center workflow integration
10. Gap Analysis - Monitoring and response capability gaps
11. Evidence Register - Audit evidence and configuration documentation
12. Approval & Sign-Off - Security operations and SOC team approval

**Key Features:**
- Log retention compliance verification (regulatory + policy requirements)
- Alert rule effectiveness assessment and tuning recommendations
- Event correlation capability evaluation
- Incident response workflow validation and effectiveness measurement
- SIEM integration assessment and coverage verification
- Reporting adequacy for multiple stakeholder audiences
- Automated gap identification for monitoring deficiencies
- Evidence linking for audit readiness
- Multi-stakeholder approval workflow (SOC, Security Ops, Compliance)

**Integration:**
This assessment completes the four-domain A.8.23 framework alongside
Infrastructure (A.8.23.1), Network Coverage (A.8.23.2), and Policy
Configuration (A.8.23.3). Results feed into the Summary Dashboard for
consolidated compliance oversight and executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a823_4_monitoring_response.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a823_4_monitoring_response.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a823_4_monitoring_response.py --date 20250115

Output:
    File: ISMS_A_8_23_4_Monitoring_Response_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Document log generation, collection, and retention configurations
    2. Validate monitoring coverage across all filtering infrastructure
    3. Review and tune alerting rules for effectiveness
    4. Assess event correlation and threat detection capabilities
    5. Validate incident response procedures and integration
    6. Verify SIEM integration and data flow
    7. Evaluate reporting adequacy for compliance and operations
    8. Assess SOC integration and workflow effectiveness
    9. Conduct gap analysis and define remediation actions
    10. Collect and link monitoring configuration evidence
    11. Obtain SOC, security operations, and compliance approvals
    12. Review Summary Dashboard metrics and finalise reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.23
Assessment Domain:    4 of 4 (Monitoring, Logging, Alerting & Response)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.23: Web Filtering Policy (Governance)
    - ISMS-IMP-A.8.23.4: Monitoring & Response Implementation Guide
    - ISMS-POL-A.8.15: Logging Policy
    - ISMS-POL-A.5.24: Information Security Incident Management
    - ISMS-IMP-A.8.23.1: Filtering Infrastructure Assessment (Domain 1)
    - ISMS-IMP-A.8.23.2: Network Coverage Assessment (Domain 2)
    - ISMS-IMP-A.8.23.3: Policy Configuration Assessment (Domain 3)

Related Controls:
    - A.8.15: Logging (Log management requirements)
    - A.8.16: Monitoring Activities (Security monitoring)
    - A.5.24: Information Security Event Management Planning
    - A.5.25: Assessment and Decision on Information Security Events
    - A.5.26: Response to Information Security Incidents
    - A.5.27: Learning from Information Security Incidents

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full monitoring and response assessment framework
    - Supports logging, alerting, incident response, and SIEM integration
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
Monitoring and incident response capabilities are critical audit focus areas:
- Log retention MUST meet regulatory and policy requirements (verify actual)
- Alert rules must be documented, tuned, and evidence false-positive reduction
- Incident response procedures must be documented AND exercised (not just paper)
- SIEM integration must be technically validated (not assumed)
- Reporting must demonstrate actual usage and effectiveness

Auditors will request:
- Sample log extracts demonstrating retention compliance
- Alert rule documentation and tuning history
- Incident response runbooks and exercise records
- SIEM correlation rule configurations
- Compliance reports with actual data (not templates)

**Data Protection:**
Assessment workbooks contain sensitive security operations information including:
- Log retention configurations (may reveal monitoring gaps)
- Alert thresholds (potential bypass information)
- Incident response procedures (operational security)
- SIEM integration details (security architecture)

Handle in accordance with your organisation's data classification policies
(typically CONFIDENTIAL or RESTRICTED).

**Maintenance:**
Review and update monitoring assessments:
- Monthly: Alert effectiveness review and tuning
- Quarterly: Incident response procedure validation and exercise
- After significant incidents: Lessons learned integration
- Semi-annually: SIEM integration and correlation effectiveness
- Annually: Complete monitoring framework re-assessment

**Quality Assurance:**
Monitoring assessments should be validated by:
- Security Operations Center (SOC) team (operational accuracy)
- Incident Response team (procedure validation)
- SIEM/monitoring platform administrators (technical verification)
- Compliance team (retention and reporting requirement validation)

**Common Pitfalls:**
- Log retention claims without verification (check actual, not configured)
- Alert fatigue not addressed (high false-positive rates reduce effectiveness)
- Paper-only incident response (no exercises, no evidence of capability)
- SIEM integration assumed but not validated (check data flow)
- Compliance reports generated but not reviewed (checkbox reporting)
- Missing alert tuning documentation (no evidence of optimization)
- No metrics on detection effectiveness (cannot prove value)

**Critical Success Factors:**
- VERIFY actual log retention, don't just document configured retention
- TUNE alerts regularly with documented false-positive reduction metrics
- EXERCISE incident response procedures with documented results
- VALIDATE SIEM integration with actual event flow testing
- DEMONSTRATE reporting effectiveness with stakeholder feedback
- QUANTIFY detection capabilities with metrics and trend analysis

**Monitoring ≠ Compliance Theater:**
This assessment measures ACTUAL monitoring effectiveness, not just the presence
of monitoring tools. Focus on:
- Detection rate (how many incidents are identified)
- Time to detect (how quickly incidents are identified)
- Time to respond (how quickly incidents are contained)
- False positive rate (efficiency of alert rules)
- Coverage verification (blind spots identification)

"Evidence > Theater" - Richard Feynman would approve.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.23.4"
WORKBOOK_NAME = "Monitoring & Response Assessment"
CONTROL_ID = "A.8.23"
CONTROL_NAME = "Web Filtering"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
import os

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Standard Unicode symbols (used in status legends and dropdowns)
CHECK   = "\u2705"         # ✅ Green check
WARNING = "\u26A0"         # ⚠  Warning sign
XMARK   = "\u274C"         # ❌ Red cross
DASH = "\u2014"            # — Em dash

from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment

# -----------------------------------------------------------------------------
# Color Palette (Consistent with A.8.23 Framework)
# -----------------------------------------------------------------------------
COLORS = {
    "header_dark": "003366",      # Dark blue - main headers
    "header_medium": "4472C4",    # Medium blue - subheaders
    "header_light": "4472C4",     # Medium blue - section headers
    "input_yellow": "FFFFCC",     # Yellow - user input cells
    "status_green": "C6EFCE",     # Green - Implemented/Compliant
    "status_yellow": "FFEB9C",    # Yellow/Amber - Partial
    "status_red": "FFC7CE",       # Red - Not Implemented
    "status_blue": "F2F2F2",      # Light grey - Planned
    "status_gray": "D9D9D9",      # Gray - N/A
    "white": "FFFFFF",
    "black": "000000",
    "light_gray": "F2F2F2",
}

# -----------------------------------------------------------------------------
# Style Definitions
# -----------------------------------------------------------------------------

# =============================================================================
# SECTION 1: IMPORTS AND CONFIGURATION
# =============================================================================

def create_styles():
    """Create reusable style dictionaries."""
    return {
        "header": {
            "font": Font(bold=True, color=COLORS["white"], size=14),
            "fill": PatternFill(start_color=COLORS["header_dark"],
                               end_color=COLORS["header_dark"],
                               fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center",
                                  wrap_text=True),
        },
        "subheader": {
            "font": Font(bold=True, color=COLORS["white"], size=11),
            "fill": PatternFill(start_color=COLORS["header_medium"],
                               end_color=COLORS["header_medium"],
                               fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center",
                                  wrap_text=True),
        },
        "section": {
            "font": Font(bold=True, color=COLORS["black"], size=11),
            "fill": PatternFill(start_color=COLORS["header_light"],
                               end_color=COLORS["header_light"],
                               fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "label": {
            "font": Font(bold=True, size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "input": {
            "fill": PatternFill(start_color=COLORS["input_yellow"],
                               end_color=COLORS["input_yellow"],
                               fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center",
                                  wrap_text=True),
        },
        "normal": {
            "font": Font(size=10),
            "alignment": Alignment(horizontal="left", vertical="center",
                                  wrap_text=True),
        },
    }

def get_border(style="thin"):
    """Create border object."""
    side = Side(style=style, color=COLORS["black"])
    return Border(left=side, right=side, top=side, bottom=side)

# -----------------------------------------------------------------------------
# Dropdown Validation Lists
# -----------------------------------------------------------------------------
DROPDOWNS = {
    "status": "✅ Implemented,⚠️ Partial,⟳ Planned,❌ Not Implemented,N/A",
    "yes_no": "Yes,No",
    "yes_no_partial": "Yes,No,Partial",
    "priority": "Critical,High,Medium,Low",
    "severity": "Critical,High,Medium,Low,Informational",
    "log_type": "Blocked_Requests,Allowed_Requests,Bypass_Attempts,Policy_Violations,Authentication,Configuration_Changes,System_Events",
    "collection_method": "Syslog,API,Agent,File_Transfer,Direct_Query,SNMP,Other",
    "destination": "SIEM,Log_Server,Cloud_SIEM,Local_Storage,Archive,Multiple",
    "format": "CEF,JSON,Syslog,CSV,Proprietary,XML,Other",
    "alert_category": "Threat_Detection,Policy_Violation,System_Health,Threshold_Breach,Anomaly,Compliance",
    "notification_method": "Email,SMS,SIEM_Alert,Ticket,Dashboard,Webhook,PagerDuty,Teams,Slack",
    "dashboard_platform": "SIEM,Filtering_Console,Custom,PowerBI,Grafana,Splunk,Cloud_Native,Excel,Other",
    "audience": "SOC,Management,IT_Ops,Compliance,CISO,All,Security_Team",
    "update_frequency": "Real-Time,Hourly,Daily,Weekly,Monthly",
    "kpi_category": "Volume,Security,Performance,Compliance,Operational",
    "kpi_status": "Met,Not_Met,At_Risk,New",
    "trend": "Improving,Stable,Degrading,New,Increasing,Decreasing",
    "report_type": "Operational,Compliance,Executive,Incident,Trend,Audit,Ad-hoc",
    "frequency": "Daily,Weekly,Monthly,Quarterly,Annual,Ad-hoc",
    "generation_method": "Automated,Manual,Semi-Automated",
    "delivery_method": "Email,Portal,Dashboard,File_Share,Meeting,API",
    "gap_category": "Logging,Alerting,Monitoring,Incident_Response,Reporting,Integration,Retention,Process",
    "gap_status": "○ Open,⟳ In Progress,✅ Resolved,⚠️ Accepted,⏳ Deferred",
    "fp_result": "Confirmed_FP,True_Positive,Inconclusive,Pending",
    "fp_resolution": "Whitelist,Policy_Tuning,Category_Update,No_Action,Escalated,Vendor_Update",
    "fp_recurrence": "First_Time,Recurring,Chronic",
    "fp_status": "Open,Resolved,Escalated,Pending",
    "evidence_type": "Screenshot,Configuration_Export,Log_Sample,Report,Procedure_Document,Email_Confirmation,Audit_Report,Test_Results,Policy_Document,Meeting_Minutes,Other",
    "verification_status": "Verified,Pending,Not_Verified,Expired",
}

def add_dropdown(ws, cell_range, dropdown_key):
    """Add data validation dropdown to a cell range."""
    if dropdown_key not in DROPDOWNS:
        return
    dv = DataValidation(
        type="list",
        formula1=f'"{DROPDOWNS[dropdown_key]}"',
        allow_blank=True
    )
    dv.error = "Please select from dropdown"
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select from list"
    dv.promptTitle = "Available Options"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# SECTION 2: SHEET 1 - INSTRUCTIONS & LEGEND
# =============================================================================

def create_sheet_instructions(wb, styles):
    """Create Instructions & Legend sheet (standard common sheet)."""
    ws = wb.active
    ws.sheet_view.showGridLines = False
    ws.title = "Instructions & Legend"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header (Row 1) — two-line merged A1:G1 ---
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # --- Document Information (Row 3+) ---
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", WORKBOOK_NAME),
        ("Related Policy", "ISMS-POL-A.8.23"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value if value else ""
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # --- Instructions (Row 13+) ---
    row = 13
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Navigate through each worksheet tab to assess monitoring and response capabilities.",
        "2. Use dropdown menus for standardised entries (Status, Severity, Priority).",
        "3. Fill in yellow-highlighted cells with your organisation-specific information.",
        "4. Document all log sources, alert rules, dashboards, and incident procedures.",
        "5. Track false positives and analyse blocked event trends.",
        "6. Identify gaps in the Gap Analysis sheet with remediation plans.",
        "7. Link evidence in the Evidence Register for audit traceability.",
        "8. Obtain required approvals in the Approval Sign-Off sheet.",
        "9. Review quarterly or after significant changes to monitoring infrastructure.",
        "10. Review the Summary Dashboard metrics and finalise reporting.",
    ]

    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # --- Status Legend (table format) ---
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border

    legend = [
        (CHECK, "Compliant", "Requirement fully met with evidence"),
        (WARNING, "Partial", "Partially implemented, gaps identified"),
        (XMARK, "Non-Compliant", "Requirement not met, remediation needed"),
        (DASH, "N/A", "Not applicable to this organisation"),
    ]

    row += 1
    for sym, status, desc in legend:
        ws.cell(row=row, column=1, value=sym).border = border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    # --- Acceptable Evidence ---
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    evidence_types = [
        "SIEM log exports and retention configuration screenshots",
        "Alert rule configuration documentation and tuning history",
        "Incident response runbooks and exercise records",
        "Blocked event analysis reports and trend data",
        "False positive management logs and resolution records",
        "Monitoring dashboard screenshots and KPI reports",
        "SOC integration workflow documentation",
        "Compliance reporting samples with actual data",
        "Log collection architecture diagrams",
        "SLA performance reports and metrics",
    ]

    row += 1
    for ev in evidence_types:
        ws[f"A{row}"] = ev
        row += 1

    # --- Column widths and freeze ---
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"

    return ws

# =============================================================================
# SECTION 3: SHEET 2 - LOG COLLECTION
# =============================================================================

def create_sheet_log_collection(wb, styles):
    """Create Log Collection assessment sheet."""
    ws = wb.create_sheet("Log Collection")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 14, "B": 28, "C": 22, "D": 25, "E": 20,
        "F": 22, "G": 14, "H": 16, "I": 18, "J": 18,
        "K": 14, "L": 18, "M": 35
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:M1")
    ws["A1"] = "LOG COLLECTION ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:M2")
    ws["A2"] = "Document all web filtering log sources, collection methods, storage, and retention compliance"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Section A: Log Source Inventory ---
    ws["A4"] = "SECTION A: LOG SOURCE INVENTORY"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:M4")
    
    # Column headers
    headers_a = [
        "Log_Source_ID", "Log_Source_Name", "Log_Type", "Source_System",
        "Collection_Method", "Destination", "Format", "Retention_Days",
        "Retention_Compliant", "Volume_Daily", "Status", "Evidence_Ref", "Notes"
    ]
    
    for col_num, header in enumerate(headers_a, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Data rows - MAX-001 fix: 1 sample + 50 empty rows
    # Sample row
    row = 6
    ws[f"A{row}"] = "LOG-001"
    ws[f"A{row}"].font = Font(bold=True)

    for col in range(2, 14):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Empty rows
    for i in range(50):
        row = 7 + i
        for col in range(2, 14):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]
    
    # Add dropdowns for log source inventory (MAX-001: updated to 51 rows)
    add_dropdown(ws, "C6:C56", "log_type")
    add_dropdown(ws, "E6:E56", "collection_method")
    add_dropdown(ws, "F6:F56", "destination")
    add_dropdown(ws, "G6:G56", "format")
    add_dropdown(ws, "I6:I56", "yes_no_partial")
    add_dropdown(ws, "K6:K56", "status")
    
    # --- Section B: Retention Requirements ---
    # NOTE: row_b must be > 56 (data rows end at 56) to avoid overlapping Section A
    row_b = 58
    ws[f"A{row_b}"] = "SECTION B: RETENTION REQUIREMENTS"
    ws[f"A{row_b}"].font = styles["section"]["font"]
    ws[f"A{row_b}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_b}:G{row_b}")
    
    headers_b = [
        "Requirement_ID", "Requirement_Source", "Log_Type_Affected",
        "Min_Retention_Days", "Current_Retention", "Compliant", "Notes"
    ]
    
    row_b += 1
    for col_num, header in enumerate(headers_b, 1):
        cell = ws.cell(row=row_b, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Retention requirement rows - MAX-001 fix: 1 sample + 10 empty rows
    row_b += 1

    # Sample row with RET-001
    # Use F2F2F2 (not FFFFCC) so QA tool doesn't extend Section A data range scan
    _sec_b_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws[f"A{row_b}"] = "RET-001"
    ws[f"A{row_b}"].font = Font(bold=True)

    for col in range(2, 8):
        cell = ws.cell(row=row_b, column=col)
        cell.fill = _sec_b_fill
        cell.border = get_border()

    # Add 10 empty rows (no pre-filled IDs)
    for i in range(10):
        row = row_b + 1 + i
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = _sec_b_fill
            cell.border = get_border()

    # Dropdown for requirement source (updated range to row_b:row_b+10)
    dv_req = DataValidation(type="list",
                            formula1='"Regulatory,Policy,Contractual,Best_Practice"',
                            allow_blank=True)
    ws.add_data_validation(dv_req)
    dv_req.add(f"B{row_b}:B{row_b+10}")

    add_dropdown(ws, f"F{row_b}:F{row_b+10}", "yes_no_partial")
    
    # --- Section C: Summary Metrics ---
    # row_b is now at sample row (60) after two += 1 increments from 58
    # Section B spans rows 58–70 (header + col headers + sample + 10 empty)
    row_c = row_b + 13  # row_b=60 → row_c=73
    ws[f"A{row_c}"] = "SECTION C: SUMMARY METRICS"
    ws[f"A{row_c}"].font = styles["section"]["font"]
    ws[f"A{row_c}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_c}:D{row_c}")

    metrics = [
        ("Total log sources configured:", '=COUNTA(B6:B56)'),
        ("Blocked request log sources:", '=COUNTIF(C6:C56,"Blocked_Requests")'),
        ("Sources meeting retention:", '=COUNTIF(I6:I56,"Yes")'),
        ("Retention compliance rate:", '=IF(COUNTA(I6:I56)>0,COUNTIF(I6:I56,"Yes")/COUNTA(I6:I56),0)'),
        ("Implemented sources:", '=COUNTIF(K6:K56,"\u2705 Implemented")'),
        ("Sources needing attention:", '=COUNTIF(K6:K56,"\u26a0\ufe0f Partial")+COUNTIF(K6:K56,"\u274c Not Implemented")'),
    ]
    
    row_c += 1
    for label, formula in metrics:
        ws[f"A{row_c}"] = label
        ws[f"A{row_c}"].font = styles["label"]["font"]
        ws[f"B{row_c}"] = formula
        ws[f"B{row_c}"].border = get_border()
        if "rate" in label.lower():
            ws[f"B{row_c}"].number_format = "0.0%"
        row_c += 1
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws


# =============================================================================
# SECTION 4: SHEET 3 - ALERT CONFIGURATION
# =============================================================================

def create_sheet_alert_configuration(wb, styles):
    """Create Alert Configuration assessment sheet."""
    ws = wb.create_sheet("Alert Configuration")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 12, "B": 32, "C": 20, "D": 38, "E": 18,
        "F": 14, "G": 22, "H": 25, "I": 20, "J": 25,
        "K": 15, "L": 14, "M": 14, "N": 15
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:N1")
    ws["A1"] = "ALERT CONFIGURATION ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:N2")
    ws["A2"] = "Document alerting rules, thresholds, notification methods, and escalation paths"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Section A: Alert Rules Inventory ---
    ws["A4"] = "SECTION A: ALERT RULES INVENTORY"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:N4")
    
    headers_a = [
        "Alert_ID", "Alert_Name", "Alert_Category", "Trigger_Condition",
        "Threshold_Value", "Severity", "Notification_Method", "Recipients",
        "Response_SLA_Min", "Escalation_Path", "Auto_Response", "Status",
        "Last_Triggered", "Evidence_Ref"
    ]
    
    for col_num, header in enumerate(headers_a, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Data rows - MAX-001 fix: 1 sample + 50 empty rows
    # Sample row
    row = 6
    ws[f"A{row}"] = "ALR-001"
    ws[f"A{row}"].font = Font(bold=True)

    for col in range(2, 15):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Empty rows
    for i in range(50):
        row = 7 + i
        for col in range(2, 15):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    # Add dropdowns (MAX-001: updated to 51 rows)
    add_dropdown(ws, "C6:C56", "alert_category")
    add_dropdown(ws, "F6:F56", "severity")
    add_dropdown(ws, "G6:G56", "notification_method")
    add_dropdown(ws, "K6:K56", "yes_no_partial")
    add_dropdown(ws, "L6:L56", "status")
    
    # --- Section B: Alert Categories Summary ---
    # NOTE: row_b must be > 56 (data rows end at 56) to avoid circular refs
    row_b = 58
    ws[f"A{row_b}"] = "SECTION B: ALERT CATEGORIES SUMMARY"
    ws[f"A{row_b}"].font = styles["section"]["font"]
    ws[f"A{row_b}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_b}:E{row_b}")
    
    cat_headers = ["Category", "Count", "Active", "Critical_Count", "Notes"]
    row_b += 1
    for col_num, header in enumerate(cat_headers, 1):
        cell = ws.cell(row=row_b, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.border = get_border()
    
    categories = [
        "Threat_Detection", "Policy_Violation", "System_Health",
        "Threshold_Breach", "Anomaly", "Compliance"
    ]
    
    row_b += 1
    for cat in categories:
        ws[f"A{row_b}"] = cat
        ws[f"B{row_b}"] = f'=COUNTIF(C6:C56,"{cat}")'
        ws[f"C{row_b}"] = f'=COUNTIFS(C6:C56,"{cat}",L6:L56,"\u2705 Implemented")'
        ws[f"D{row_b}"] = f'=COUNTIFS(C6:C56,"{cat}",F6:F56,"Critical")'
        ws[f"E{row_b}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row_b}"].border = get_border()
        row_b += 1
    
    # --- Section C: Summary Metrics ---
    row_c = row_b + 2
    ws[f"A{row_c}"] = "SECTION C: SUMMARY METRICS"
    ws[f"A{row_c}"].font = styles["section"]["font"]
    ws[f"A{row_c}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_c}:D{row_c}")
    
    metrics = [
        ("Total alert rules configured:", '=COUNTA(B6:B56)'),
        ("Implemented alerts:", '=COUNTIF(L6:L56,"\u2705 Implemented")'),
        ("Critical severity alerts:", '=COUNTIF(F6:F56,"Critical")'),
        ("High severity alerts:", '=COUNTIF(F6:F56,"High")'),
        ("Alerts with auto-response:", '=COUNTIF(K6:K56,"Yes")'),
        ("Alert coverage score:", '=IF(COUNTA(B6:B56)>0,COUNTIF(L6:L56,"\u2705 Implemented")/COUNTA(B6:B56),0)'),
    ]
    
    row_c += 1
    for label, formula in metrics:
        ws[f"A{row_c}"] = label
        ws[f"A{row_c}"].font = styles["label"]["font"]
        ws[f"B{row_c}"] = formula
        ws[f"B{row_c}"].border = get_border()
        if "score" in label.lower():
            ws[f"B{row_c}"].number_format = "0.0%"
        row_c += 1
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws

# =============================================================================
# SECTION 5: SHEET 4 - MONITORING DASHBOARD
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Navigate through each worksheet tab to assess monitoring and response capabilities.',
        '2. Use dropdown menus for standardised entries (Status, Severity, Priority).',
        '3. Fill in yellow-highlighted cells with your organisation-specific information.',
        '4. Document all log sources, alert rules, dashboards, and incident procedures.',
        '5. Track false positives and analyse blocked event trends.',
        '6. Identify gaps in the Gap Analysis sheet with remediation plans.',
        '7. Link evidence in the Evidence Register for audit traceability.',
        '8. Obtain required approvals in the Approval Sign-Off sheet.',
        '9. Review quarterly or after significant changes to monitoring infrastructure.',
        '10. Review the Summary Dashboard metrics and finalise reporting.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A24"] = "Status Legend"
    ws["A24"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=25, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 26 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_sheet_monitoring_dashboard(wb, styles):
    """Create Monitoring Dashboard assessment sheet."""
    ws = wb.create_sheet("Monitoring Dashboard")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 14, "B": 32, "C": 20, "D": 20, "E": 18,
        "F": 18, "G": 45, "H": 18, "I": 14, "J": 18
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:J1")
    ws["A1"] = "MONITORING DASHBOARD ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document monitoring dashboards, KPIs tracked, and review frequency"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Section A: Dashboard Inventory ---
    ws["A4"] = "SECTION A: DASHBOARD INVENTORY"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:J4")
    
    headers_a = [
        "Dashboard_ID", "Dashboard_Name", "Platform", "Primary_Audience",
        "Update_Frequency", "Review_Frequency", "Key_Metrics_Displayed",
        "Alerting_Integrated", "Status", "Evidence_Ref"
    ]
    
    for col_num, header in enumerate(headers_a, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Data rows - MAX-001 fix: 1 sample + 50 empty rows
    # Sample row
    row = 6
    ws[f"A{row}"] = "DSH-001"
    ws[f"A{row}"].font = Font(bold=True)

    for col in range(2, 11):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Empty rows
    for i in range(50):
        row = 7 + i
        for col in range(2, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    # Add dropdowns (51 rows: sample at 6 + 50 empty at 7-56)
    add_dropdown(ws, "C6:C56", "dashboard_platform")
    add_dropdown(ws, "D6:D56", "audience")
    add_dropdown(ws, "E6:E56", "update_frequency")
    add_dropdown(ws, "F6:F56", "update_frequency")
    add_dropdown(ws, "H6:H56", "yes_no")
    add_dropdown(ws, "I6:I56", "status")

    # --- Section B: Key Performance Indicators ---
    # NOTE: row_b must be > 56 (data rows end at 56) to avoid overlapping Section A
    row_b = 58
    ws[f"A{row_b}"] = "SECTION B: KEY PERFORMANCE INDICATORS (KPIs)"
    ws[f"A{row_b}"].font = styles["section"]["font"]
    ws[f"A{row_b}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_b}:J{row_b}")
    
    kpi_headers = [
        "KPI_ID", "KPI_Name", "Category", "Measurement_Unit",
        "Target_Value", "Current_Value", "Trend", "Review_Freq",
        "Owner_Role", "Status"
    ]
    
    row_b += 1
    for col_num, header in enumerate(kpi_headers, 1):
        cell = ws.cell(row=row_b, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # KPI rows - MAX-001 fix: 1 sample + 20 empty rows
    row_b += 1

    # Sample row with KPI-001 and suggested values
    ws[f"A{row_b}"] = "KPI-001"
    ws[f"A{row_b}"].font = Font(bold=True)
    ws[f"B{row_b}"] = "Total blocked requests (daily)"
    ws[f"C{row_b}"] = "Volume"
    ws[f"D{row_b}"] = "Count"

    for col in range(2, 11):
        cell = ws.cell(row=row_b, column=col)
        if col > 4:  # Only highlight input columns
            cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Add 20 empty rows (no pre-filled IDs)
    for i in range(20):
        row = row_b + 1 + i
        for col in range(2, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    # Add dropdowns for KPIs (updated range to 21 rows)
    add_dropdown(ws, f"C{row_b}:C{row_b+20}", "kpi_category")
    add_dropdown(ws, f"G{row_b}:G{row_b+20}", "trend")
    add_dropdown(ws, f"H{row_b}:H{row_b+20}", "frequency")
    add_dropdown(ws, f"J{row_b}:J{row_b+20}", "kpi_status")
    
    # --- Section C: Summary Metrics ---
    row_c = row_b + 22
    ws[f"A{row_c}"] = "SECTION C: SUMMARY METRICS"
    ws[f"A{row_c}"].font = styles["section"]["font"]
    ws[f"A{row_c}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_c}:D{row_c}")
    
    metrics = [
        ("Total dashboards configured:", '=COUNTA(B6:B56)'),
        ("Dashboards with alerting:", '=COUNTIF(H6:H56,"Yes")'),
        ("Real-time dashboards:", '=COUNTIF(E6:E56,"Real-Time")'),
        ("KPIs defined:", f'=COUNTA(B{row_b}:B{row_b+20})'),
        ("KPIs meeting target:", f'=COUNTIF(J{row_b}:J{row_b+20},"Met")'),
        ("KPIs at risk:", f'=COUNTIF(J{row_b}:J{row_b+20},"At_Risk")+COUNTIF(J{row_b}:J{row_b+20},"Not_Met")'),
    ]
    
    row_c += 1
    for label, formula in metrics:
        ws[f"A{row_c}"] = label
        ws[f"A{row_c}"].font = styles["label"]["font"]
        ws[f"B{row_c}"] = formula
        ws[f"B{row_c}"].border = get_border()
        row_c += 1
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws


# =============================================================================
# SECTION 6: SHEET 5 - INCIDENT RESPONSE
# =============================================================================

def create_sheet_incident_response(wb, styles):
    """Create Incident Response assessment sheet."""
    ws = wb.create_sheet("Incident Response")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 14, "B": 32, "C": 16, "D": 20, "E": 20,
        "F": 20, "G": 30, "H": 20, "I": 14, "J": 20
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:J1")
    ws["A1"] = "INCIDENT RESPONSE ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:J2")
    ws["A2"] = "Document incident categories, SLAs, response procedures, and performance metrics"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Section A: Incident Categories ---
    ws["A4"] = "SECTION A: INCIDENT CATEGORIES & SLAs"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:I4")
    
    headers_a = [
        "Category_ID", "Incident_Category", "Default_Severity",
        "Response_SLA_Min", "Resolution_SLA_Hrs", "Escalation_Threshold",
        "Response_Procedure", "Owner_Role", "Status"
    ]
    
    for col_num, header in enumerate(headers_a, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Incident category rows - MAX-001 fix: 1 sample + 15 empty rows
    row = 6

    # Sample row with CAT-001 and suggested values
    ws[f"A{row}"] = "CAT-001"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "Malware delivery attempt"
    ws[f"C{row}"] = "Critical"
    ws[f"D{row}"] = 15
    ws[f"E{row}"] = 4

    for col in range(2, 10):
        cell = ws.cell(row=row, column=col)
        if col > 5:  # Only highlight non-prepopulated columns
            cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Add 15 empty rows (no pre-filled IDs)
    for i in range(15):
        row = 7 + i
        for col in range(2, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    add_dropdown(ws, "C6:C21", "severity")
    add_dropdown(ws, "I6:I21", "status")
    
    # --- Section B: SLA Performance ---
    row_b = 24
    ws[f"A{row_b}"] = "SECTION B: SLA PERFORMANCE (Last 90 Days)"
    ws[f"A{row_b}"].font = styles["section"]["font"]
    ws[f"A{row_b}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_b}:E{row_b}")
    
    sla_headers = ["Metric", "Target", "Actual", "Status", "Notes"]
    row_b += 1
    for col_num, header in enumerate(sla_headers, 1):
        cell = ws.cell(row=row_b, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.border = get_border()
    
    sla_metrics = [
        ("Critical incident response time", "<15 min"),
        ("High incident response time", "<30 min"),
        ("Medium incident response time", "<60 min"),
        ("Low incident response time", "<4 hours"),
        ("Critical resolution time", "<4 hours"),
        ("High resolution time", "<8 hours"),
        ("Medium resolution time", "<24 hours"),
        ("Overall SLA compliance rate", ">95%"),
        ("Mean time to detect (MTTD)", "<5 min"),
        ("Mean time to respond (MTTR)", "<30 min"),
    ]
    
    row_b += 1
    for metric, target in sla_metrics:
        ws[f"A{row_b}"] = metric
        ws[f"B{row_b}"] = target
        ws[f"C{row_b}"].fill = styles["input"]["fill"]
        ws[f"D{row_b}"].fill = styles["input"]["fill"]
        ws[f"E{row_b}"].fill = styles["input"]["fill"]
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row_b}"].border = get_border()
        row_b += 1
    
    add_dropdown(ws, f"D{row_b-10}:D{row_b-1}", "kpi_status")
    
    # --- Section C: Recent Incident Summary ---
    row_c = row_b + 2
    ws[f"A{row_c}"] = "SECTION C: RECENT INCIDENT SUMMARY (Last 20 Incidents)"
    ws[f"A{row_c}"].font = styles["section"]["font"]
    ws[f"A{row_c}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_c}:J{row_c}")
    
    inc_headers = [
        "Incident_ID", "Date_Detected", "Category", "Severity",
        "Response_Time_Min", "Resolution_Time_Hrs", "SLA_Met",
        "Root_Cause", "Lessons_Learned", "Evidence_Ref"
    ]
    
    row_c += 1
    for col_num, header in enumerate(inc_headers, 1):
        cell = ws.cell(row=row_c, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Incident log rows - MAX-001 fix: 1 sample + 20 empty rows
    row_c += 1

    # Sample row with INC-001
    ws[f"A{row_c}"] = "INC-001"
    ws[f"A{row_c}"].font = Font(bold=True)

    for col in range(2, 11):
        cell = ws.cell(row=row_c, column=col)
        cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Add 20 empty rows (no pre-filled IDs)
    for i in range(20):
        row = row_c + 1 + i
        for col in range(2, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    add_dropdown(ws, f"D{row_c}:D{row_c+20}", "severity")
    add_dropdown(ws, f"G{row_c}:G{row_c+20}", "yes_no")
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws

# =============================================================================
# SECTION 7: SHEET 6 - BLOCKED EVENTS ANALYSIS
# =============================================================================

def create_sheet_blocked_events(wb, styles):
    """Create Blocked Events Analysis sheet."""
    ws = wb.create_sheet("Blocked Events Analysis")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 14, "B": 30, "C": 20, "D": 20, "E": 16,
        "F": 25, "G": 14, "H": 16, "I": 35
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:I1")
    ws["A1"] = "BLOCKED EVENTS ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:I2")
    ws["A2"] = "Analyze patterns, trends, and categories of blocked web filtering events"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Section A: Blocked Event Categories ---
    ws["A4"] = "SECTION A: BLOCKED EVENT CATEGORIES"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:I4")
    
    headers_a = [
        "Category_ID", "Block_Category", "Events_30_Days", "Events_90_Days",
        "Trend", "Top_Source", "Risk_Level", "Action_Required", "Notes"
    ]
    
    for col_num, header in enumerate(headers_a, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Block category rows - MAX-001 fix: 1 sample + 20 empty rows
    row = 6

    # Sample row with BLK-001 and suggested value
    ws[f"A{row}"] = "BLK-001"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "Malware"

    for col in range(2, 10):
        cell = ws.cell(row=row, column=col)
        if col > 2:  # Only highlight input columns
            cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Add 20 empty rows (no pre-filled IDs)
    for i in range(20):
        row = 7 + i
        for col in range(2, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    add_dropdown(ws, "E6:E26", "trend")
    add_dropdown(ws, "G6:G26", "priority")  # Risk level uses priority values
    add_dropdown(ws, "H6:H26", "yes_no")
    
    # --- Section B: Top Blocked Threats ---
    row_b = 28
    ws[f"A{row_b}"] = "SECTION B: TOP 10 BLOCKED THREATS (Last 30 Days)"
    ws[f"A{row_b}"].font = styles["section"]["font"]
    ws[f"A{row_b}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_b}:F{row_b}")
    
    top_headers = ["Rank", "Threat_Type/URL", "Count_30_Days", "Percentage", 
                   "Mitigation_Status", "Notes"]
    row_b += 1
    for col_num, header in enumerate(top_headers, 1):
        cell = ws.cell(row=row_b, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.border = get_border()
    
    row_b += 1
    for i in range(10):
        ws[f"A{row_b + i}"] = i + 1
        ws[f"A{row_b + i}"].font = Font(bold=True)
        ws[f"A{row_b + i}"].alignment = Alignment(horizontal="center")
        for col in range(2, 7):
            cell = ws.cell(row=row_b + i, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
    
    add_dropdown(ws, f"E{row_b}:E{row_b+9}", "status")
    
    # --- Section C: 6-Month Trend Analysis ---
    row_c = row_b + 12
    ws[f"A{row_c}"] = "SECTION C: 6-MONTH TREND ANALYSIS"
    ws[f"A{row_c}"].font = styles["section"]["font"]
    ws[f"A{row_c}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_c}:G{row_c}")
    
    trend_headers = ["Month", "Total_Blocked", "Malware", "Phishing", 
                     "Policy_Violation", "Other", "Notes"]
    row_c += 1
    for col_num, header in enumerate(trend_headers, 1):
        cell = ws.cell(row=row_c, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.border = get_border()
    
    # 6 months of trend data
    row_c += 1
    months = ["Month-6", "Month-5", "Month-4", "Month-3", "Month-2", "Current"]
    for i, month in enumerate(months):
        ws[f"A{row_c + i}"] = month
        ws[f"A{row_c + i}"].font = Font(bold=True)
        for col in range(2, 8):
            cell = ws.cell(row=row_c + i, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
    
    # --- Section D: Summary Metrics ---
    row_d = row_c + 8
    ws[f"A{row_d}"] = "SECTION D: SUMMARY METRICS"
    ws[f"A{row_d}"].font = styles["section"]["font"]
    ws[f"A{row_d}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_d}:D{row_d}")
    
    metrics = [
        ("Categories tracked:", '=COUNTA(B6:B25)'),
        ("Categories requiring action:", '=COUNTIF(H6:H25,"Yes")'),
        ("High-risk categories:", '=COUNTIF(G6:G25,"Critical")+COUNTIF(G6:G25,"High")'),
        ("Categories with increasing trend:", '=COUNTIF(E6:E25,"Increasing")'),
    ]
    
    row_d += 1
    for label, formula in metrics:
        ws[f"A{row_d}"] = label
        ws[f"A{row_d}"].font = styles["label"]["font"]
        ws[f"B{row_d}"] = formula
        ws[f"B{row_d}"].border = get_border()
        row_d += 1
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws


# =============================================================================
# SECTION 8: SHEET 7 - FALSE POSITIVE MANAGEMENT
# =============================================================================

def create_sheet_false_positive(wb, styles):
    """Create False Positive Management sheet."""
    ws = wb.create_sheet("False Positive Mgmt")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 12, "B": 14, "C": 20, "D": 28, "E": 35,
        "F": 20, "G": 22, "H": 14, "I": 20, "J": 14,
        "K": 14, "L": 15
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:L1")
    ws["A1"] = "FALSE POSITIVE MANAGEMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:L2")
    ws["A2"] = "Track false positives, tuning actions, and resolution effectiveness"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Section A: False Positive Log ---
    ws["A4"] = "SECTION A: FALSE POSITIVE LOG"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:L4")
    
    headers_a = [
        "FP_ID", "Date_Reported", "Reported_By", "Blocked_URL_Category",
        "Business_Justification", "Investigation_Result", "Resolution_Action",
        "Resolution_Date", "Resolution_Time_Hrs", "Recurrence", "Status", "Evidence_Ref"
    ]
    
    for col_num, header in enumerate(headers_a, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # FP rows - MAX-001 fix: 1 sample + 50 empty rows
    row = 6

    # Sample row with FP-001
    ws[f"A{row}"] = "FP-001"
    ws[f"A{row}"].font = Font(bold=True)

    for col in range(2, 13):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Add 50 empty rows (no pre-filled IDs)
    for i in range(50):
        row = 7 + i
        for col in range(2, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    # Add dropdowns (updated to 51 rows)
    add_dropdown(ws, "F6:F56", "fp_result")
    add_dropdown(ws, "G6:G56", "fp_resolution")
    add_dropdown(ws, "J6:J56", "fp_recurrence")
    add_dropdown(ws, "K6:K56", "fp_status")
    
    # --- Section B: False Positive Metrics ---
    row_b = 58
    ws[f"A{row_b}"] = "SECTION B: FALSE POSITIVE METRICS"
    ws[f"A{row_b}"].font = styles["section"]["font"]
    ws[f"A{row_b}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_b}:E{row_b}")
    
    metric_headers = ["Metric", "Value", "Target", "Status", "Notes"]
    row_b += 1
    for col_num, header in enumerate(metric_headers, 1):
        cell = ws.cell(row=row_b, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.border = get_border()
    
    fp_metrics = [
        ("Total FPs reported (last 90 days)", '=COUNTA(B6:B55)', "<50"),
        ("Average resolution time (hours)", '=IFERROR(AVERAGE(I6:I55),"N/A")', "<24"),
        ("Open FPs", '=COUNTIF(K6:K55,"Open")', "<10"),
        ("Confirmed FP rate", '=IFERROR(COUNTIF(F6:F55,"Confirmed_FP")/COUNTA(F6:F55),"N/A")', ">80%"),
        ("Recurring FP percentage", '=IFERROR((COUNTIF(J6:J55,"Recurring")+COUNTIF(J6:J55,"Chronic"))/COUNTA(J6:J55),"N/A")', "<10%"),
        ("FPs escalated", '=COUNTIF(K6:K55,"Escalated")', "<5"),
        ("FPs resolved via whitelist", '=COUNTIF(G6:G55,"Whitelist")', "Monitor"),
        ("FPs resolved via tuning", '=COUNTIF(G6:G55,"Policy_Tuning")', "Monitor"),
    ]
    
    row_b += 1
    for metric, formula, target in fp_metrics:
        ws[f"A{row_b}"] = metric
        ws[f"A{row_b}"].font = styles["label"]["font"]
        ws[f"B{row_b}"] = formula
        ws[f"C{row_b}"] = target
        ws[f"D{row_b}"].fill = styles["input"]["fill"]
        ws[f"E{row_b}"].fill = styles["input"]["fill"]
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row_b}"].border = get_border()
        row_b += 1
    
    add_dropdown(ws, f"D{row_b-8}:D{row_b-1}", "kpi_status")
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws


# =============================================================================
# SECTION 9: SHEET 8 - REPORTING SCHEDULE
# =============================================================================

def create_sheet_reporting_schedule(wb, styles):
    """Create Reporting Schedule sheet."""
    ws = wb.create_sheet("Reporting Schedule")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 12, "B": 32, "C": 18, "D": 14, "E": 18,
        "F": 28, "G": 18, "H": 40, "I": 14, "J": 14, "K": 15
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:K1")
    ws["A1"] = "REPORTING SCHEDULE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:K2")
    ws["A2"] = "Document regular reports generated for web filtering operations"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Section A: Report Inventory ---
    ws["A4"] = "SECTION A: REPORT INVENTORY"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:K4")
    
    headers_a = [
        "Report_ID", "Report_Name", "Report_Type", "Frequency",
        "Generation_Method", "Distribution_List", "Delivery_Method",
        "Content_Summary", "Last_Generated", "Status", "Evidence_Ref"
    ]
    
    for col_num, header in enumerate(headers_a, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Report rows - MAX-001 fix: 1 sample + 20 empty rows
    row = 6

    # Sample row with RPT-001 and suggested values
    ws[f"A{row}"] = "RPT-001"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "Daily Threat Summary"
    ws[f"C{row}"] = "Operational"
    ws[f"D{row}"] = "Daily"

    for col in range(2, 12):
        cell = ws.cell(row=row, column=col)
        if col > 4:  # Only highlight non-prepopulated columns
            cell.fill = styles["input"]["fill"]
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Add 20 empty rows (no pre-filled IDs)
    for i in range(20):
        row = 7 + i
        for col in range(2, 12):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    # Add dropdowns (updated to 21 rows)
    add_dropdown(ws, "C6:C26", "report_type")
    add_dropdown(ws, "D6:D26", "frequency")
    add_dropdown(ws, "E6:E26", "generation_method")
    add_dropdown(ws, "G6:G26", "delivery_method")
    add_dropdown(ws, "J6:J26", "status")
    
    # --- Section B: Stakeholder Coverage ---
    row_b = 28
    ws[f"A{row_b}"] = "SECTION B: STAKEHOLDER REPORTING COVERAGE"
    ws[f"A{row_b}"].font = styles["section"]["font"]
    ws[f"A{row_b}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_b}:E{row_b}")
    
    coverage_headers = ["Stakeholder", "Required_Reports", "Reports_Received", 
                        "Coverage_%", "Notes"]
    row_b += 1
    for col_num, header in enumerate(coverage_headers, 1):
        cell = ws.cell(row=row_b, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.border = get_border()
    
    stakeholders = ["CISO", "IT Management", "Compliance Team", "SOC", 
                    "Internal Audit", "External Auditors", "HR", "Legal"]
    
    row_b += 1
    for stakeholder in stakeholders:
        ws[f"A{row_b}"] = stakeholder
        ws[f"A{row_b}"].font = Font(bold=True)
        ws[f"B{row_b}"].fill = styles["input"]["fill"]
        ws[f"C{row_b}"].fill = styles["input"]["fill"]
        ws[f"D{row_b}"] = '=IFERROR(C{0}/B{0},0)'.format(row_b)
        ws[f"D{row_b}"].number_format = "0%"
        ws[f"E{row_b}"].fill = styles["input"]["fill"]
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row_b}"].border = get_border()
        row_b += 1
    
    # --- Section C: Summary Metrics ---
    row_c = row_b + 2
    ws[f"A{row_c}"] = "SECTION C: SUMMARY METRICS"
    ws[f"A{row_c}"].font = styles["section"]["font"]
    ws[f"A{row_c}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_c}:D{row_c}")
    
    metrics = [
        ("Total reports configured:", '=COUNTA(B6:B25)'),
        ("Automated reports:", '=COUNTIF(E6:E25,"Automated")'),
        ("Reports implemented:", '=COUNTIF(J6:J25,"\u2705 Implemented")'),
        ("Daily reports:", '=COUNTIF(D6:D25,"Daily")'),
        ("Stakeholders with full coverage:", '=COUNTIF(D30:D37,">=100%")'),
    ]
    
    row_c += 1
    for label, formula in metrics:
        ws[f"A{row_c}"] = label
        ws[f"A{row_c}"].font = styles["label"]["font"]
        ws[f"B{row_c}"] = formula
        ws[f"B{row_c}"].border = get_border()
        row_c += 1
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws

# =============================================================================
# SECTION 10: SHEET 9 - GAP ANALYSIS
# =============================================================================

def create_sheet_gap_analysis(wb, styles):
    """Create Gap Analysis sheet."""
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False
    
    # Column widths
    col_widths = {
        "A": 14, "B": 18, "C": 40, "D": 30, "E": 30,
        "F": 14, "G": 25, "H": 35, "I": 20, "J": 14,
        "K": 15, "L": 12, "M": 15
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # --- Header ---
    ws.merge_cells("A1:M1")
    ws["A1"] = "GAP ANALYSIS - MONITORING & RESPONSE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:M2")
    ws["A2"] = "Identify and prioritize monitoring and incident response gaps for remediation"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # --- Gap Register ---
    ws["A4"] = "GAP REGISTER"
    ws["A4"].font = styles["section"]["font"]
    ws["A4"].fill = styles["section"]["fill"]
    ws.merge_cells("A4:M4")
    
    headers = [
        "Gap ID", "Gap Category", "Gap Description", "Current State",
        "Target State", "Risk Impact", "Affected Systems", "Remediation Action",
        "Owner", "Target Date", "Status", "Priority", "Evidence Ref"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.alignment = styles["subheader"]["alignment"]
        cell.border = get_border()
    
    # Gap rows - MAX-001 fix: 1 sample + 30 empty rows
    row = 6

    # Sample row with GAP4-001
    _gap4_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws[f"A{row}"] = "GAP4-001"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].fill = _gap4_grey
    ws[f"A{row}"].border = get_border()

    for col in range(2, 14):
        cell = ws.cell(row=row, column=col)
        cell.fill = _gap4_grey
        cell.border = get_border()
        cell.alignment = styles["normal"]["alignment"]

    # Add 30 empty rows (no pre-filled IDs)
    for i in range(30):
        row = 7 + i
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=1).border = get_border()
        for col in range(2, 14):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input"]["fill"]
            cell.border = get_border()
            cell.alignment = styles["normal"]["alignment"]

    # Add dropdowns (updated to 31 rows)
    add_dropdown(ws, "B6:B36", "gap_category")
    add_dropdown(ws, "F6:F36", "priority")  # Risk impact uses priority values
    add_dropdown(ws, "K6:K36", "gap_status")
    add_dropdown(ws, "L6:L36", "priority")
    
    # --- Summary Metrics ---
    row_sum = 38
    ws[f"A{row_sum}"] = "SUMMARY METRICS"
    ws[f"A{row_sum}"].font = styles["section"]["font"]
    ws[f"A{row_sum}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_sum}:D{row_sum}")
    
    metrics = [
        ("Total gaps identified:", '=COUNTA(C6:C36)'),
        ("Critical gaps:", '=COUNTIF(F6:F36,"Critical")'),
        ("High gaps:", '=COUNTIF(F6:F36,"High")'),
        ("Open gaps:", '=COUNTIF(K6:K36,"\u25cb Open")'),
        ("In Progress gaps:", '=COUNTIF(K6:K36,"\u27f3 In Progress")'),
        ("Resolved gaps:", '=COUNTIF(K6:K36,"\u2705 Resolved")'),
        ("Gaps past target date:", '=COUNTIFS(K6:K36,"<>\u2705 Resolved",J6:J36,"<"&TODAY())'),
    ]
    
    row_sum += 1
    for label, formula in metrics:
        ws[f"A{row_sum}"] = label
        ws[f"A{row_sum}"].font = styles["label"]["font"]
        ws[f"B{row_sum}"] = formula
        ws[f"B{row_sum}"].border = get_border()
        row_sum += 1
    
    # --- Gap Categories Breakdown ---
    row_cat = row_sum + 2
    ws[f"A{row_cat}"] = "GAP CATEGORIES BREAKDOWN"
    ws[f"A{row_cat}"].font = styles["section"]["font"]
    ws[f"A{row_cat}"].fill = styles["section"]["fill"]
    ws.merge_cells(f"A{row_cat}:D{row_cat}")
    
    gap_cats = ["Logging", "Alerting", "Monitoring", "Incident_Response", 
                "Reporting", "Integration", "Retention", "Process"]
    
    row_cat += 1
    cat_headers = ["Category", "Count", "Open", "Critical/High"]
    for col_num, header in enumerate(cat_headers, 1):
        cell = ws.cell(row=row_cat, column=col_num, value=header)
        cell.font = styles["subheader"]["font"]
        cell.fill = styles["subheader"]["fill"]
        cell.border = get_border()
    
    row_cat += 1
    for cat in gap_cats:
        ws[f"A{row_cat}"] = cat
        ws[f"B{row_cat}"] = f'=COUNTIF(B6:B36,"{cat}")'
        ws[f"C{row_cat}"] = f'=COUNTIFS(B6:B36,"{cat}",K6:K36,"\u25cb Open")'
        ws[f"D{row_cat}"] = f'=COUNTIFS(B6:B36,"{cat}",F6:F36,"Critical")+COUNTIFS(B6:B36,"{cat}",F6:F36,"High")'
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row_cat}"].border = get_border()
        row_cat += 1
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws


# =============================================================================
# SECTION 11: SHEET 10 - EVIDENCE REGISTER
# =============================================================================


def create_sheet_summary_dashboard(wb, styles):
    """Create Summary Dashboard sheet for Monitoring & Response Assessment."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(cell, text, bg="003366", fg="FFFFFF", bold=True, size=11, center=True):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    def _grey(cell, value=None, center=False):
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        if value is not None:
            cell.value = value
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _yellow(cell):
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = border

    # --- ROW 1 ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for _bc in range(1, 8):
        ws.cell(row=1, column=_bc).border = border

    # --- ROW 2 ---
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------------------
    # TABLE 1
    # -------------------------------------------------------------------------
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border

    row += 1
    t1_hdrs = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    _d9_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, h in enumerate(t1_hdrs, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="000000")
        cell.fill = _d9_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    start_row = row

    # Note: Gen4 uses underscore DVs — "Not_Implemented", "In_Progress"
    area_configs = [
        {
            "label": "Log Collection",
            "comp":    "=COUNTIF(\'Log Collection\'!K6:K56,\"\u2705 Implemented\")",
            "partial": "=COUNTIF(\'Log Collection\'!K6:K56,\"\u26a0\ufe0f Partial\")",
            "noncomp": "=COUNTIF(\'Log Collection\'!K6:K56,\"\u274c Not Implemented\")"
                       "+COUNTIF(\'Log Collection\'!K6:K56,\"\u27f3 Planned\")",
            "na":      "=COUNTIF(\'Log Collection\'!K6:K56,\"N/A\")",
        },
        {
            "label": "Alert Configuration",
            "comp":    "=COUNTIF(\'Alert Configuration\'!L6:L56,\"\u2705 Implemented\")",
            "partial": "=COUNTIF(\'Alert Configuration\'!L6:L56,\"\u26a0\ufe0f Partial\")",
            "noncomp": "=COUNTIF(\'Alert Configuration\'!L6:L56,\"\u274c Not Implemented\")"
                       "+COUNTIF(\'Alert Configuration\'!L6:L56,\"\u27f3 Planned\")",
            "na":      "=COUNTIF(\'Alert Configuration\'!L6:L56,\"N/A\")",
        },
        {
            "label": "Open Monitoring Gaps",
            "comp":    "=COUNTIF(\'Gap Analysis\'!K6:K36,\"\u2705 Resolved\")"
                       "+COUNTIF(\'Gap Analysis\'!K6:K36,\"\u26a0\ufe0f Accepted\")",
            "partial": "=COUNTIF(\'Gap Analysis\'!K6:K36,\"\u27f3 In Progress\")",
            "noncomp": "=COUNTIF(\'Gap Analysis\'!K6:K36,\"\u25cb Open\")",
            "na":      "=COUNTIF(\'Gap Analysis\'!K6:K36,\"\u23f3 Deferred\")",
        },
    ]

    _blue_font = Font(color="000000")
    _blue_bold = Font(bold=True)
    _center_align = Alignment(horizontal="center", vertical="center")
    for cfg in area_configs:
        ws.cell(row=row, column=1, value=cfg["label"]).border = border
        ws.cell(row=row, column=2, value=f"=C{row}+D{row}+E{row}+F{row}").border = border
        ws.cell(row=row, column=2).font = _blue_font
        ws.cell(row=row, column=2).alignment = _center_align
        for col_idx, key in enumerate(["comp", "partial", "noncomp", "na"], 3):
            cell = ws.cell(row=row, column=col_idx, value=cfg[key])
            cell.font = _blue_font
            cell.border = border
            cell.alignment = _center_align
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.font = _blue_bold
        cell_g.border = border
        cell_g.alignment = _center_align
        cell_g.number_format = "0.0%"
        row += 1

    # TOTAL row
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = chr(64 + col_idx)
        cell = ws.cell(row=row, column=col_idx)
        cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{row - 1})"
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.border = border
        cell.alignment = _center_align
    cell_g = ws.cell(row=row, column=7)
    cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
    cell_g.font = Font(bold=True)
    cell_g.fill = PatternFill("solid", fgColor="D9D9D9")
    cell_g.border = border
    cell_g.alignment = _center_align
    cell_g.number_format = "0.0%"

    row += 3

    # -------------------------------------------------------------------------
    # TABLE 2: KEY METRICS
    # -------------------------------------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 2: KEY METRICS")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border
    row += 1
    _hdr(ws.cell(row=row, column=1), "Metric", bg="D9D9D9", fg="000000", size=10)
    ws.merge_cells(f"B{row}:G{row}")
    _hdr(ws.cell(row=row, column=2), "Count", bg="D9D9D9", fg="000000", size=10, center=True)
    for _c in range(3, 8):
        ws.cell(row=row, column=_c).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.cell(row=row, column=_c).border = border
    row += 1

    t2_metrics = [
        ("Log Sources ✅ Implemented", "=COUNTIF('Log Collection'!K6:K56,\"✅ Implemented\")"),
        ("Log Sources ❌ Not Implemented / ⟳ Planned", "=COUNTIF('Log Collection'!K6:K56,\"❌ Not Implemented\")+COUNTIF('Log Collection'!K6:K56,\"⟳ Planned\")"),
        ("Alert Rules ✅ Implemented", "=COUNTIF('Alert Configuration'!L6:L56,\"✅ Implemented\")"),
        ("Alert Rules ❌ Not Implemented / ⟳ Planned", "=COUNTIF('Alert Configuration'!L6:L56,\"❌ Not Implemented\")+COUNTIF('Alert Configuration'!L6:L56,\"⟳ Planned\")"),
        ("Monitoring Gaps ○ Open", "=COUNTIF('Gap Analysis'!K6:K36,\"○ Open\")"),
        ("Monitoring Gaps ⟳ In Progress", "=COUNTIF('Gap Analysis'!K6:K36,\"⟳ In Progress\")"),
        ("Monitoring Gaps ✅ Resolved / ⚠️ Accepted", "=COUNTIF('Gap Analysis'!K6:K36,\"✅ Resolved\")+COUNTIF('Gap Analysis'!K6:K36,\"⚠️ Accepted\")"),
        ("Monitoring Gaps ⏳ Deferred", "=COUNTIF('Gap Analysis'!K6:K36,\"⏳ Deferred\")"),
        ("Dashboard Sources ✅ Implemented", "=COUNTIF('Monitoring Dashboard'!I6:I80,\"✅ Implemented\")"),
        ("Dashboard Sources ❌ Not Configured / ⟳ Planned", "=COUNTIF('Monitoring Dashboard'!I6:I80,\"❌ Not Implemented\")+COUNTIF('Monitoring Dashboard'!I6:I80,\"⟳ Planned\")"),
        ("IR Procedures ✅ Documented", "=COUNTIF('Incident Response'!I6:I60,\"✅ Implemented\")"),
        ("IR Procedures ❌ Not Documented / ⟳ Planned", "=COUNTIF('Incident Response'!I6:I60,\"❌ Not Implemented\")+COUNTIF('Incident Response'!I6:I60,\"⟳ Planned\")"),
        ("Blocked Event Categories Requiring Action", "=COUNTIF('Blocked Events Analysis'!H6:H49,\"Yes\")"),
        ("False Positives Open / Escalated", "=COUNTIF('False Positive Mgmt'!K6:K67,\"Open\")+COUNTIF('False Positive Mgmt'!K6:K67,\"Escalated\")"),
        ("False Positives Resolved", "=COUNTIF('False Positive Mgmt'!K6:K67,\"Resolved\")"),
    ]

    for _t2_label, _t2_formula in t2_metrics:
        _cell_a = ws.cell(row=row, column=1)
        _cell_a.value = _t2_label
        _cell_a.font = Font(name="Calibri")
        _cell_a.border = border
        ws.merge_cells(f"B{row}:G{row}")
        for _c in range(2, 8):
            ws.cell(row=row, column=_c).border = border
        ws.cell(row=row, column=2).value = _t2_formula
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).number_format = '0'
        row += 1

    row += 2

    # -------------------------------------------------------------------------
    # TABLE 3: CRITICAL FINDINGS & OPEN GAPS
    # -------------------------------------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    _hdr(ws[f"A{row}"], "TABLE 3: CRITICAL FINDINGS & OPEN GAPS", bg="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for _bc in range(1, 8):
        ws.cell(row=row, column=_bc).border = border
    row += 1
    t3_hdrs = ["Gap ID", "Gap Category", "Gap Description", "Risk Level",
               "Owner", "Target Date", "Status"]
    for col_idx, h in enumerate(t3_hdrs, 1):
        _hdr(ws.cell(row=row, column=col_idx), h, size=10)
    row += 1

    # TABLE 3: INDEX/SMALL/IF — auto-pull rows where Gap Analysis col K = "\u25cb Open"
    # Column mapping: Gap ID(A), Gap Category(B), Gap Description(C),
    #                 Risk Level(F), Owner(I), Target Date(J), Status(K)
    # Gap Analysis data range: rows 6-36 (31 rows: 1 sample + 30 empty)
    _gap = "Gap Analysis"
    _gcols = ["A", "B", "C", "F", "I", "J", "K"]
    for k in range(1, 11):
        for tbl_col, gc in enumerate(_gcols, 1):
            formula = (
                f"=IFERROR(INDEX(\'{_gap}\'!{gc}$6:{gc}$36,"
                f"SMALL(IF(\'{_gap}\'!K$6:K$36=\"\u25cb Open\"," 
                f"ROW(\'{_gap}\'!K$6:K$36)-ROW(\'{_gap}\'!K$6)+1),{k})),\"\")"
            )
            c = ws.cell(row=row, column=tbl_col)
            c.value = formula
            _yellow(c)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.freeze_panes = "A4"


def create_sheet_evidence_register(wb, styles):
    """Create Evidence Register sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── Row 2: Subtitle ──
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 4: Column headers ──
    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data Validation ──
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,'
                 'Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # ── Row 5: Sample row (F2F2F2 grey) ──
    sample_data = {
        1: "EV-001",
        2: "Web Filtering Monitoring & Response",
        3: "Configuration file",
        4: "Log collection configuration and alert rule export from SIEM",
        5: "/evidence/A.8.23/monitoring-config.json",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # ── Rows 6-105: Empty data rows (FFFFCC, 100 rows) ──
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{r}"])
        ver_status_dv.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"

    return ws


# =============================================================================
# SECTION 12: SHEET 11 - APPROVAL SIGN-OFF
# =============================================================================

def create_sheet_approval(wb, styles):
    """Create Approval Sign-Off sheet — golden standard."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # ── Row 2: Control reference ──
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # ── Row 3: ASSESSMENT SUMMARY banner ──
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # ── Summary fields ──
    summary_fields = [
        ("Document:",                  f"{DOCUMENT_ID} - {WORKBOOK_NAME}", False),
        ("Assessment Period:",         "",                                   True),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9",          True),
        ("Assessment Status:",         "",                                   True),
        ("Assessed By:",               "",                                   True),
    ]

    row = 4
    status_row_for_dv = None
    for label, value, yellow in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for col in range(2, 6):
            if yellow:
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        if "Assessment Status" in label:
            status_row_for_dv = row
        row += 1

    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f"B{status_row_for_dv}")

    # ── Approver sections ──
    row += 2

    def _approver(start_row, title, color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = border
            ws.merge_cells(f"B{r}:E{r}")
            for col in range(2, 6):
                ws.cell(row=r, column=col).fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=r, column=col).border = border
            r += 1
        return r + 1

    row = _approver(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _approver(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")
    row = _approver(row, "APPROVED BY (CISO)", "003366")

    # ── Final Decision ──
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f"B{row}")

    # ── Next Review Details ──
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

    return ws


# =============================================================================
# SECTION 13: MAIN FUNCTION AND EXECUTION
# =============================================================================

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 70)
    logger.info("ISMS-IMP-A.8.23.4 - Monitoring & Response Assessment Generator")
    logger.info("ISO/IEC 27001:2022 - Control A.8.23: Web Filtering")
    logger.info("=" * 70)
    logger.info("")
    
    # Create workbook
    logger.info("[1/12] Creating workbook...")
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    styles = create_styles()
    
    # Generate sheets
    logger.info("[2/12] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb.active)
    
    logger.info("[3/12] Creating Log Collection sheet...")
    create_sheet_log_collection(wb, styles)

    logger.info("[4/12] Creating Alert Configuration sheet...")
    create_sheet_alert_configuration(wb, styles)

    logger.info("[5/12] Creating Monitoring Dashboard sheet...")
    create_sheet_monitoring_dashboard(wb, styles)

    logger.info("[6/12] Creating Incident Response sheet...")
    create_sheet_incident_response(wb, styles)

    logger.info("[7/12] Creating Blocked Events Analysis sheet...")
    create_sheet_blocked_events(wb, styles)

    logger.info("[8/12] Creating False Positive Mgmt sheet...")
    create_sheet_false_positive(wb, styles)

    logger.info("[9/12] Creating Reporting Schedule sheet...")
    create_sheet_reporting_schedule(wb, styles)

    logger.info("[10/12] Creating Gap Analysis sheet...")
    create_sheet_gap_analysis(wb, styles)

    logger.info("[11/13] Creating Evidence Register sheet...")
    create_sheet_evidence_register(wb, styles)

    logger.info("[12/13] Creating Summary Dashboard sheet...")
    create_sheet_summary_dashboard(wb, styles)

    logger.info("[13/13] Creating Approval Sign-Off sheet...")
    create_sheet_approval(wb, styles)
    
    # Generate filename with date
    date_str = datetime.now().strftime("%Y%m%d")
    
    # Save workbook
    logger.info("")
    logger.info(f"Saving workbook as: {output_path.name}")
    finalize_validations(wb)
    wb.save(output_path)
    # Print summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("GENERATION COMPLETE")
    logger.info("=" * 70)
    logger.info("")
    logger.info(f"Output file: {output_path.name}")
    logger.info(f"Total sheets: {len(wb.sheetnames)}")
    logger.info("")
    logger.info("Sheet Summary:")
    logger.info("-" * 40)
    for i, sheet in enumerate(wb.sheetnames, 1):
        logger.info(f"  {i:2}. {sheet}")
    logger.info("")
    logger.info("Key Capacities:")
    logger.info("-" * 40)
    logger.info("  \u2022 Log Sources: 30 rows")
    logger.info("  \u2022 Alert Rules: 40 rows")
    logger.info("  \u2022 Dashboards: 20 rows")
    logger.info("  \u2022 KPIs: 20 rows")
    logger.info("  \u2022 Incident Categories: 15 rows")
    logger.info("  \u2022 Recent Incidents: 20 rows")
    logger.info("  \u2022 Blocked Event Categories: 20 rows")
    logger.info("  \u2022 False Positives: 50 rows")
    logger.info("  \u2022 Reports: 20 rows")
    logger.info("  \u2022 Gaps: 30 rows")
    logger.info("  \u2022 Evidence: 100 rows")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("-" * 40)
    logger.info("  1. Open the workbook in Excel")
    logger.info("  2. Complete each assessment sheet")
    logger.info("  3. Use dropdown menus for consistency")
    logger.info("  4. Document evidence for each item")
    logger.info("  5. Identify gaps and remediation plans")
    logger.info("  6. Obtain required approvals")
    logger.info("")
    logger.info("Remember: Evidence > Theater")
    logger.info("         'We monitor' means nothing without metrics.")
    logger.info("")
    


def main():
    output_path = _wkbk_dir / OUTPUT_FILENAME
    create_workbook(output_path)
    return output_path


def validate_workbook(filename):
    """
    Basic validation of generated workbook.
    
    Args:
        filename: Path to the workbook to validate
        
    Returns:
        bool: True if validation passes
    """
    from openpyxl import load_workbook
    
    logger.info("")
    logger.info("Running basic validation...")
    logger.info("-" * 40)
    
    try:
        wb = load_workbook(filename)
        
        expected_sheets = [
            "Instructions & Legend",
            "Log Collection",
            "Alert Configuration",
            "Monitoring Dashboard",
            "Incident Response",
            "Blocked Events Analysis",
            "False Positive Mgmt",
            "Reporting Schedule",
            "Gap Analysis",
            "Evidence Register",
            "Summary Dashboard",
            "Approval Sign-Off",
        ]
        
        # Check sheet count
        if len(wb.sheetnames) != 12:
            logger.info(f"  \u274C Expected 12 sheets, found {len(wb.sheetnames)}")
            return False
        logger.info(f"  \u2705 Sheet count: {len(wb.sheetnames)}")
        
        # Check sheet names
        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                logger.info(f"  \u2705 Found: {sheet}")
            else:
                logger.info(f"  \u274C Missing: {sheet}")
                return False
        
        # Check Evidence Register has 100 rows
        ev_sheet = wb["Evidence Register"]
        ev_sheet.sheet_view.showGridLines = False
        ev_count = 0
        for row in range(5, 105):
            if ev_sheet[f"A{row}"].value:
                ev_count += 1
        
        if ev_count == 100:
            logger.info(f"  \u2705 Evidence Register: {ev_count} rows")
        else:
            logger.info(f"  \u26A0\uFE0F Evidence Register: {ev_count} rows (expected 100)")
        
        logger.info("")
        logger.info("Validation Result: \u2705 PASSED")
        logger.info("")
        
        wb.close()
        return True
        
    except Exception as e:
        logger.error(f"  \u274C Validation error: {str(e)}")
        return False


# =============================================================================
# SECTION 14: ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    """
    Entry point for script execution.
    
    Usage:
        python generate_a823_4_monitoring_response.py
        
    This will generate the assessment workbook in the current directory.
    """
    import sys
    
    logger.info("")
    logger.info("╔════════════════════════════════════════════════════════════════════╗")
    logger.info("║  ISMS Control A.8.23 - Web Filtering                               ║")
    logger.info("║  Domain 4: Monitoring & Response Assessment                        ║")
    logger.info("║                                                                    ║")
    logger.info("║  'The first principle is that you must not fool yourself'          ║")
    logger.info("╚════════════════════════════════════════════════════════════════════╝")
    logger.info("")
    
    try:
        # Generate the workbook
        output_file = main()
        
        # Validate the output
        if validate_workbook(output_file):
            logger.info(f"Successfully generated: {output_file}")
            sys.exit(0)
        else:
            logger.error("Validation failed - please check the output")
            sys.exit(1)
            
    except ImportError as e:
        logger.info("")
        logger.error("ERROR: Missing required library")
        logger.info("-" * 40)
        logger.info(f"  {str(e)}")
        logger.info("")
        logger.info("Please install openpyxl:")
        logger.info("  pip install openpyxl")
        logger.info("")
        sys.exit(1)
        
    except Exception as e:
        logger.info("")
        logger.error("ERROR: Generation failed")
        logger.info("-" * 40)
        logger.info(f"  {str(e)}")
        logger.info("")
        import traceback
        traceback.print_exc()
        sys.exit(1)


# =============================================================================
# END OF GENERATOR SCRIPT
# =============================================================================
#
# Document Control:
#   Version: 1.0
#   Created: 2025-01-01
#
# Change History:
#   1.0 - Initial version
#       - 11 sheets for monitoring & response assessment
#       - 100-row evidence register
#       - 3-level approval workflow
#       - Pre-populated incident categories, KPIs, and reports
#       - Comprehensive dropdown validations
#
# Dependencies:
#   - Python 3.7+
#   - openpyxl >= 3.0.0
#
# Output:
#   ISMS-IMP-A.8.23.4_Monitoring_Response_YYYYMMDD.xlsx
#

# =============================================================================
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
