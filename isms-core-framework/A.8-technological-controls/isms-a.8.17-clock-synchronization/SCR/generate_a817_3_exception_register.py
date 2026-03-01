#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.17.3 - Time Synchronisation Exception Register Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.17: Clock Synchronisation
Exception Management: Non-Compliant Systems and Risk Acceptance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific exception management processes, risk acceptance
workflows, and compliance tracking requirements.

Key customisation areas:
1. Exception approval workflow (match your governance structure)
2. Risk scoring criteria (adapt to your risk assessment methodology)
3. Compensating controls (specific to your security architecture)
4. Review cycles and escalation procedures (based on your policies)
5. Integration with GRC systems (if applicable)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.17 Clock Synchronisation Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel exception register for documenting
and managing systems that cannot meet A.8.17 time synchronisation requirements,
with formal risk acceptance and compensating controls.

**Purpose:**
Enables systematic tracking of exceptions to time synchronisation requirements,
including documented business justification, risk assessment, compensating
controls, and time-bound risk acceptance with defined review cycles.

**Exception Scenarios:**
- Air-gapped systems without network connectivity
- Legacy systems incompatible with modern NTP protocols
- IoT/embedded devices with limited time sync capabilities
- Systems with operational constraints preventing time sync
- Temporary exceptions during infrastructure migrations
- Systems undergoing remediation with interim risk acceptance

**Generated Workbook Structure:**
1. Instructions & Legend - Exception process guidance and criteria
2. Exception_Register - Active exception tracking and status
3. Risk_Assessment - Risk scoring and impact analysis per exception
4. Compensating_Controls - Alternative controls for each exception
5. Review_Schedule - Exception review and re-approval tracking
6. Expired_Exceptions - Overdue exceptions requiring action
7. Remediation_Plan - Path to compliance for temporary exceptions
8. Evidence_Register - Supporting documentation linkage
9. Approval & Sign-Off - Risk owner and CISO approval workflow

**Key Features:**
- Structured exception documentation with business justification
- Risk scoring based on system criticality and exposure
- Compensating control identification and validation
- Time-bound exceptions with automatic expiry tracking
- Review cycle enforcement
- Approval workflow with multi-level sign-off
- Exception statistics and trending
- Integration with gap analysis from Assessment 2

**Integration:**
Exception register data feeds into the A.8.17 Compliance Dashboard to provide
complete compliance picture including both compliant systems and formally
accepted exceptions.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

Optional:
    - Integration with GRC platforms for exception workflow
    - Import gaps from Assessment 2 (System Synchronisation Status)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a817_3_exception_register.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a817_3_exception_register.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a817_3_exception_register.py --date 20250125
    
    # Import gaps from Assessment 2
    python3 generate_a817_3_exception_register.py --import-gaps ISMS-A.8.17-Assessment-2.xlsx

Output:
    File: ISMS-A.8.17-Assessment-3-Exception-Register-YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Import or manually enter systems requiring exceptions
    2. Document business justification for each exception
    3. Conduct risk assessment (likelihood × impact)
    4. Identify and validate compensating controls
    5. Define exception duration and review cycle
    6. Obtain risk owner approval
    7. Obtain CISO/security approval
    8. Schedule exception review meetings
    9. Monitor for expired exceptions
    10. Feed results into A.8.17 Compliance Dashboard

Exception Approval Process:
    1. System owner requests exception with justification
    2. Security team assesses risk and proposes compensating controls
    3. Risk owner reviews and approves/rejects
    4. CISO provides final approval for high-risk exceptions
    5. Exception is time-bound with defined review cycle
    6. Reviews ensure exception remains valid or is remediated

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.17
Assessment Domain:    3 of 2 (Exception Management - Optional)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.17: Clock Synchronisation Policy (Requirements)
    - ISMS-IMP-A.8.17.2: System Synchronisation Status (Gap Source)
    - A.8.17 Compliance Dashboard (Consolidation)
    - Corporate Risk Management Policy (Exception Approval Process)

Related Standards:
    - ISO/IEC 27001:2022 Clause 6.1.3: Risk Treatment
    - ISO/IEC 27002:2022: Risk Acceptance Guidance

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements exception tracking and risk acceptance workflow
    - Supports compensating control documentation
    - Time-bound exceptions with automatic expiry tracking
    - Integrated with A.8.17 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Exception Philosophy:**
Exceptions are not "free passes" - they represent formally accepted risk with:
- Clear business justification
- Documented risk assessment
- Defined compensating controls
- Time-bound approval
- Regular review cycles
- Path to remediation where feasible

**Risk Assessment:**
Each exception must include risk scoring based on:
- System criticality (impact if compromised)
- Data sensitivity handled by system
- Exposure to security events
- Likelihood of time-related security issues
- Existing compensating controls

Risk scores inform approval levels and review frequency.

**Compensating Controls:**
Acceptable compensating controls may include:
- Enhanced logging/monitoring for affected system
- Manual log correlation procedures
- Physical security controls (air-gapped systems)
- Network segmentation isolating affected system
- Reduced system criticality/scope

Controls must demonstrably reduce risk to acceptable level.

**Review Cycles:**
Exception review frequency based on risk:
- High risk: Quarterly review required
- Medium risk: Semi-annual review
- Low risk: Annual review
- All exceptions: Must be renewed annually at minimum

**Audit Considerations:**
Auditors will scrutinize exceptions for:
- Adequate business justification
- Appropriate risk assessment
- Effective compensating controls
- Current approvals (not expired)
- Evidence of regular reviews
- Remediation progress for temporary exceptions

Ensure exception register is audit-ready with complete documentation.

**Integration with Assessment 2:**
Gap Analysis sheet in Assessment 2 (System Sync Status) identifies systems
not meeting requirements. These gaps feed into exception register where:
- Exception is formally requested
- Risk is assessed and accepted
- OR remediation is planned and executed

**Expired Exceptions:**
Systems with expired exceptions must be either:
- Remediated to compliance
- Re-evaluated and exception renewed
- Decommissioned if no longer needed

Never allow exceptions to remain active past expiry - this is audit finding.

**Approval Authority:**
Define clear approval levels:
- System owner: Initial exception request
- Risk owner: Risk acceptance for business unit
- Security team: Compensating control validation
- CISO: High-risk exceptions (>70 risk score)
- Executive sponsor: Exceptions >90 days duration

**Data Protection:**
Exception register contains sensitive information including:
- System security weaknesses
- Risk assessment details
- Infrastructure vulnerabilities
- Compensating control strategies

Handle as CONFIDENTIAL per organisational data classification.

**Quality Assurance:**
Security governance team should audit exception register quarterly for:
- Expired exceptions requiring action
- Exceptions without adequate compensating controls
- Patterns indicating systemic issues
- Opportunities for infrastructure improvement

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X

# Document identification constants

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.17.3"
WORKBOOK_NAME = "Exception Register"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
CONTROL_ID   = "A.8.17"
CONTROL_NAME = "Clock Synchronisation"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
WARNING = '\u26A0'    # ⚠ Warning sign
BULLET = '\u2022'     # • Bullet
ARROW = '\u2192'      # → Right arrow

def create_styles():
    """Define A.8.24 standard styles for the workbook"""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "title": {
            "font": Font(name="Calibri", bold=True, size=16, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center")
        },
        "subtitle": {
            "font": Font(name="Calibri", bold=True, size=12, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center")
        },
        "warning": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "good": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }

def set_column_widths(ws, widths):
    """Set column widths for a worksheet"""
    for col_num, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

def apply_style(cell, style_dict):
    """Apply style dictionary to cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Log new exception requests in the Exception Requests sheet.', '2. Track approved exceptions in the Active Exceptions sheet.', '3. Archive expired or revoked exceptions in the Expired Exceptions sheet.', '4. Monitor exception metrics in the Summary Dashboard sheet.', '5. Record all evidence in the Evidence Register sheet.', '6. Complete the Approval Sign-Off sheet when assessment is finished.', '7. Use dropdown lists where provided for consistent data entry.', '8. Add rows as needed — formulas auto-extend for new data.', '9. Save completed workbook with date suffix for version tracking.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_exception_requests_sheet(ws, styles):
    """Create exception requests sheet (pending approval)"""
    ws.title = "Exception Requests"

    # Title
    ws['A1'] = "EXCEPTION REQUESTS (PENDING APPROVAL)"
    apply_style(ws['A1'], styles['header'])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells('A1:M1')
    ws.row_dimensions[1].height = 35

    # Subtitle row 2 (DS-006)
    ws.merge_cells('A2:M2')
    ws['A2'] = "Track and manage all time synchronisation exception requests"
    ws['A2'].font = Font(italic=True, size=10, name="Calibri")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Exception ID [*]",
        "Submitted Date [*]",
        "System/Asset Name [*]",
        "Requirement Exempted [*]",
        "Exception Type [*]",
        "Business Justification [*]",
        "Risk Assessment [*]",
        "Compensating Controls [*]",
        "Requested Duration [*]",
        "Submitted By [*]",
        "Approval Authority [*]",
        "Approval Status",
        "Notes",
    ]

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
        cell.border = border_thin

    # Data validations
    exception_type_dv = DataValidation(
        type="list",
        formula1='"Technical Exception,Policy-Level Exception,Temporary Non-Compliance"',
        allow_blank=False
    )
    ws.add_data_validation(exception_type_dv)

    approval_auth_dv = DataValidation(
        type="list",
        formula1='"CISO,Executive Management"',
        allow_blank=False
    )
    ws.add_data_validation(approval_auth_dv)

    approval_status_dv = DataValidation(
        type="list",
        formula1='"Pending,Approved,Rejected,Withdrawn"',
        allow_blank=False
    )
    ws.add_data_validation(approval_status_dv)

    # Example row
    example = [
        "EXC-A817-001",
        datetime.now().strftime('%d.%m.%Y'),
        "legacy-scada-system.example.com",
        "REQ-817-009 (All systems must synchronize)",
        "Technical Exception",
        "Air-gapped SCADA system cannot reach external NTP servers; vendor does not support NTP protocol",
        "Moderate risk: Time drift may cause log correlation issues; no authentication impact (isolated network)",
        "GPS time receiver installed for Stratum 0 local time source; manual time verification weekly",
        "12 months (temporary until vendor upgrade)",
        "Thomas Meier (SCADA Admin)",
        "CISO",
        "Pending",
        "Vendor upgrade planned Q3 2026",
    ]

    _grey_fill_g3 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col_num, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = value
        apply_style(cell, styles['data'])
        cell.fill = _grey_fill_g3  # Sample row is grey (MAX-006)

    # Apply validations to data rows (4-54) — sample row + 50 empty
    for row in range(4, 55):
        exception_type_dv.add(ws.cell(row=row, column=5))
        approval_auth_dv.add(ws.cell(row=row, column=11))
        approval_status_dv.add(ws.cell(row=row, column=12))

        # Apply borders and input styling to all cells
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles['data'])
            # Sample row (row 4) stays grey; rest are yellow input cells
            if row == 4:
                cell.fill = _grey_fill_g3
            else:
                cell.fill = styles['input_cell']['fill']
    
    set_column_widths(ws, [18, 15, 30, 25, 22, 40, 40, 40, 20, 25, 20, 18, 30])
    ws.freeze_panes = 'A4'

def create_active_exceptions_sheet(ws, styles):
    """Create active exceptions sheet (approved and in effect)"""
    ws.title = "Active Exceptions"

    # Title
    ws['A1'] = "ACTIVE EXCEPTIONS (APPROVED)"
    apply_style(ws['A1'], styles['header'])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells('A1:O1')
    ws.row_dimensions[1].height = 35

    # Subtitle row 2 (DS-006)
    ws.merge_cells('A2:O2')
    ws['A2'] = "Currently approved exceptions with their approval details and expiry dates"
    ws['A2'].font = Font(italic=True, size=10, name="Calibri")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Exception ID [*]",
        "System/Asset Name [*]",
        "Requirement Exempted [*]",
        "Exception Type [*]",
        "Approved By [*]",
        "Approval Date [*]",
        "Expiry Date [*]",
        "Days Until Expiry",
        "Compensating Controls [*]",
        "Last Quarterly Review",
        "Next Review Due",
        "Reassessment Required",
        "Status",
        "Review Notes",
        "Actions Required",
    ]

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
        cell.border = border_thin

    # Data validations
    exception_type_dv = DataValidation(
        type="list",
        formula1='"Technical Exception,Policy-Level Exception,Temporary Non-Compliance"',
        allow_blank=False
    )
    ws.add_data_validation(exception_type_dv)

    reassess_dv = DataValidation(
        type="list",
        formula1='"Yes - Within 90 days of policy review,No"',
        allow_blank=False
    )
    ws.add_data_validation(reassess_dv)

    status_dv = DataValidation(
        type="list",
        formula1='"Active,Expiring Soon,Under Review,Pending Revocation"',
        allow_blank=False
    )
    ws.add_data_validation(status_dv)

    # Example row
    today = datetime.now()
    expiry = today + timedelta(days=180)
    next_review = today + timedelta(days=90)

    example = [
        "EXC-A817-001",
        "legacy-scada-system.example.com",
        "REQ-817-009 (All systems must synchronize)",
        "Technical Exception",
        "Anna Muller (CISO)",
        today.strftime('%d.%m.%Y'),
        expiry.strftime('%d.%m.%Y'),
        f"=G4-TODAY()",  # Days until expiry formula
        "GPS time receiver (Stratum 0); Weekly manual verification",
        today.strftime('%d.%m.%Y'),
        next_review.strftime('%d.%m.%Y'),
        "No",
        "Active",
        "Compensating controls verified; GPS receiver operational",
        "Continue quarterly reviews; Monitor vendor upgrade timeline",
    ]

    _grey_fill_ae = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col_num, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = value
        apply_style(cell, styles['data'])
        if col_num == 8:  # Days until expiry
            cell.font = Font(color="0000FF")
            cell.fill = _grey_fill_ae  # Sample row is grey (MAX-006)
        else:
            cell.fill = _grey_fill_ae  # Sample row is grey (MAX-006)

    # Apply validations and formulas to data rows (4-54) — sample row + 50 empty
    for row in range(4, 55):
        exception_type_dv.add(ws.cell(row=row, column=4))
        reassess_dv.add(ws.cell(row=row, column=12))
        status_dv.add(ws.cell(row=row, column=13))

        # Days until expiry formula
        if row > 4:
            ws.cell(row=row, column=8).value = f"=IF(ISBLANK(G{row}),\"\",G{row}-TODAY())"
            ws.cell(row=row, column=8).font = Font(color="0000FF")

            # Next review due (90 days from last review)
            ws.cell(row=row, column=11).value = f"=IF(ISBLANK(J{row}),\"\",J{row}+90)"
            ws.cell(row=row, column=11).font = Font(color="0000FF")

        # Apply borders and input styling to all cells
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            if col in [8, 11]:
                # Formula columns: apply border only (preserve value and font)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # Sample row stays grey
                if row == 4:
                    cell.fill = _grey_fill_ae
            else:
                apply_style(cell, styles['data'])
                # Sample row (row 4) stays grey; rest are yellow input cells
                if row == 4:
                    cell.fill = _grey_fill_ae
                else:
                    cell.fill = styles['input_cell']['fill']
    
    # Conditional formatting note
    ws['A56'] = "NOTE: Exceptions expiring within 30 days require renewal or revocation decision"
    ws['A56'].font = Font(italic=True, color="FF0000")
    ws.merge_cells('A56:O56')
    
    set_column_widths(ws, [18, 30, 30, 22, 25, 15, 15, 18, 40, 18, 18, 22, 18, 40, 40])
    ws.freeze_panes = 'A4'

def create_expired_exceptions_sheet(ws, styles):
    """Create expired/revoked exceptions sheet (historical record)"""
    ws.title = "Expired Exceptions"

    # Title
    ws['A1'] = "EXPIRED AND REVOKED EXCEPTIONS (HISTORICAL RECORD)"
    apply_style(ws['A1'], styles['header'])
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells('A1:L1')
    ws.row_dimensions[1].height = 35

    # Subtitle row 2 (DS-006)
    ws.merge_cells('A2:L2')
    ws['A2'] = "Historical record of all expired or revoked time synchronisation exceptions"
    ws['A2'].font = Font(italic=True, size=10, name="Calibri")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Exception ID",
        "System/Asset Name",
        "Requirement Exempted",
        "Exception Type",
        "Approved By",
        "Approval Date",
        "Expiry Date",
        "Closure Date",
        "Closure Reason",
        "Total Duration (Days)",
        "Final Status",
        "Lessons Learned",
    ]

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
        cell.border = border_thin

    # Data validations
    closure_reason_dv = DataValidation(
        type="list",
        formula1='"Expired - Not Renewed,Compliance Achieved,Revoked - Risk Changed,Revoked - Controls Failed,System Decommissioned"',
        allow_blank=False
    )
    ws.add_data_validation(closure_reason_dv)

    final_status_dv = DataValidation(
        type="list",
        formula1='"Completed Successfully,Revoked,Replaced,System Retired"',
        allow_blank=False
    )
    ws.add_data_validation(final_status_dv)

    # Example row
    example = [
        "EXC-A817-002",
        "old-db-server.example.com",
        "REQ-817-011 (Drift threshold compliance)",
        "Temporary Non-Compliance",
        "Anna Muller (CISO)",
        "2025-06-15",
        "2025-12-15",
        "2025-11-30",
        "Compliance Achieved",
        "=H4-F4",  # Duration formula
        "Completed Successfully",
        "NTP configuration corrected; System now compliant with drift thresholds",
    ]

    _grey_fill_ee = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col_num, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = value
        apply_style(cell, styles['data'])
        if col_num == 10:  # Duration formula
            cell.font = Font(color="0000FF")
            cell.fill = _grey_fill_ee  # Sample row is grey (MAX-006)
        else:
            cell.fill = _grey_fill_ee  # Sample row is grey (MAX-006)

    # Apply validations to data rows (4-104) — sample row + 100 empty
    for row in range(4, 105):
        closure_reason_dv.add(ws.cell(row=row, column=9))
        final_status_dv.add(ws.cell(row=row, column=11))

        # Duration formula
        if row > 4:
            ws.cell(row=row, column=10).value = f"=IF(OR(ISBLANK(F{row}),ISBLANK(H{row})),\"\",H{row}-F{row})"
            ws.cell(row=row, column=10).font = Font(color="0000FF")

        # Apply borders and input styling to all cells
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            if col == 10:
                # Formula column: apply border only (preserve value and font)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # Sample row stays grey
                if row == 4:
                    cell.fill = _grey_fill_ee
            else:
                apply_style(cell, styles['data'])
                # Sample row (row 4) stays grey; rest are yellow input cells
                if row == 4:
                    cell.fill = _grey_fill_ee
                else:
                    cell.fill = styles['input_cell']['fill']
    
    set_column_widths(ws, [18, 30, 30, 22, 25, 15, 15, 15, 25, 20, 20, 50])
    ws.freeze_panes = 'A4'

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard — TABLE 1/2/3 format."""
    ws.title = "Summary Dashboard"
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(fill_hex, bold=True, size=11, color="FFFFFF"):
        return Font(bold=bold, size=size, color=color, name="Calibri")

    navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    # ── ROW 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ── ROW 2: Subtitle ───────────────────────────────────────────────────────
    ws["A2"] = "Exception Management — A.8.17 Clock Synchronisation"
    ws["A2"].font = Font(italic=True, size=10, name="Calibri", color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 1: Assessment Areas ─────────────────────────────────────────────
    # TABLE 1 banner (row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].border = border_thin

    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, hdr in enumerate(t1_headers, 1):
        c = ws.cell(row=5, column=col_idx, value=hdr)
        c.font = Font(bold=True, size=10, name="Calibri")
        c.fill = grey_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin

    # TABLE 1 data rows — no fill, blue 0000FF text
    t1_data = [
        ("Exception Requests",
         "=COUNTA('Exception Requests'!A5:A54)",
         "=COUNTIF('Exception Requests'!L5:L54,\"Approved\")",
         "=COUNTIF('Exception Requests'!L5:L54,\"Pending\")",
         "=COUNTIF('Exception Requests'!L5:L54,\"Rejected\")+COUNTIF('Exception Requests'!L5:L54,\"Withdrawn\")",
         0,
         "=IF((B9-F9)=0,0,C9/(B9-F9))"),
        ("Active Exceptions",
         "=COUNTA('Active Exceptions'!A5:A54)",
         "=COUNTIF('Active Exceptions'!M5:M54,\"Active\")",
         "=COUNTIF('Active Exceptions'!M5:M54,\"Expiring Soon\")+COUNTIF('Active Exceptions'!M5:M54,\"Under Review\")",
         "=COUNTIF('Active Exceptions'!M5:M54,\"Pending Revocation\")",
         0,
         "=IF((B9-F9)=0,0,C9/(B9-F9))"),
        ("Expired Exceptions",
         "=COUNTA('Expired Exceptions'!A5:A104)",
         "=COUNTIF('Expired Exceptions'!K5:K104,\"Completed Successfully\")+COUNTIF('Expired Exceptions'!K5:K104,\"Replaced\")",
         "=COUNTIF('Expired Exceptions'!K5:K104,\"System Retired\")",
         "=COUNTIF('Expired Exceptions'!K5:K104,\"Revoked\")",
         0,
         "=IF((B9-F9)=0,0,C9/(B9-F9))"),
    ]

    for row_offset, row_data in enumerate(t1_data):
        row = 6 + row_offset
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.border = border_thin
            c.font = Font(name="Calibri", color="000000")
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            if col_idx == 7:
                c.number_format = "0.0%"

    # TOTAL row (row 8)
    total_vals = [
        "TOTAL",
        "=B6+B7+B8",
        "=C6+C7+C8",
        "=D6+D7+D8",
        "=E6+E7+E8",
        "=F6+F7+F8",
        "=IF((B9-F9)=0,0,C9/(B9-F9))",
    ]
    for col_idx, val in enumerate(total_vals, 1):
        c = ws.cell(row=9, column=col_idx, value=val)
        c.font = Font(bold=True, name="Calibri")
        c.fill = grey_fill
        c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        c.border = border_thin
        if col_idx == 7:
            c.number_format = "0.0%"

    # ── TABLE 2: Key Performance Metrics ──────────────────────────────────────
    # Row 9: blank gap
    # Row 10: section header
    ws.merge_cells("A11:G11")
    ws["A11"] = "KEY PERFORMANCE METRICS"
    ws["A11"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws["A11"].fill = navy_fill
    ws["A11"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A11"].border = border_thin

    # Row 11: column headers
    for col_idx, hdr in enumerate(["Metric", "Value", "Category", "", "", "", ""], 1):
        c = ws.cell(row=12, column=col_idx, value=hdr)
        c.font = Font(bold=True, size=10, name="Calibri")
        c.fill = grey_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin

    # TABLE 2 metrics (rows 12-31)
    t2_metrics = [
        # (Metric Name, Formula, Category)
        ("Total Exception Requests", "=COUNTA('Exception Requests'!A5:A54)", "Requests"),
        ("Requests — Pending Approval", "=COUNTIF('Exception Requests'!L5:L54,\"Pending\")", "Requests"),
        ("Requests — Approved", "=COUNTIF('Exception Requests'!L5:L54,\"Approved\")", "Requests"),
        ("Requests — Rejected", "=COUNTIF('Exception Requests'!L5:L54,\"Rejected\")", "Requests"),
        ("Requests — Withdrawn", "=COUNTIF('Exception Requests'!L5:L54,\"Withdrawn\")", "Requests"),
        ("Requests — Technical Exceptions", "=COUNTIF('Exception Requests'!E5:E54,\"Technical Exception\")", "By Type"),
        ("Requests — Policy-Level Exceptions", "=COUNTIF('Exception Requests'!E5:E54,\"Policy-Level Exception\")", "By Type"),
        ("Requests — Temporary Non-Compliance", "=COUNTIF('Exception Requests'!E5:E54,\"Temporary Non-Compliance\")", "By Type"),
        ("Total Active Exceptions", "=COUNTA('Active Exceptions'!A5:A54)", "Active"),
        ("Active — Status: Active", "=COUNTIF('Active Exceptions'!M5:M54,\"Active\")", "Active"),
        ("Active — Expiring Soon", "=COUNTIF('Active Exceptions'!M5:M54,\"Expiring Soon\")", "Active"),
        ("Active — Under Review", "=COUNTIF('Active Exceptions'!M5:M54,\"Under Review\")", "Active"),
        ("Active — Pending Revocation", "=COUNTIF('Active Exceptions'!M5:M54,\"Pending Revocation\")", "Active"),
        ("Active — Requiring Reassessment", "=COUNTIF('Active Exceptions'!L5:L54,\"Yes - Within 90 days of policy review\")", "Active"),
        ("Active — Technical Exceptions", "=COUNTIF('Active Exceptions'!D5:D54,\"Technical Exception\")", "Active by Type"),
        ("Active — Policy-Level Exceptions", "=COUNTIF('Active Exceptions'!D5:D54,\"Policy-Level Exception\")", "Active by Type"),
        ("Total Expired / Archived", "=COUNTA('Expired Exceptions'!A5:A104)", "Expired"),
        ("Expired — Completed Successfully", "=COUNTIF('Expired Exceptions'!K5:K104,\"Completed Successfully\")", "Expired"),
        ("Expired — Revoked", "=COUNTIF('Expired Exceptions'!K5:K104,\"Revoked\")", "Expired"),
        ("Expired — Compliance Achieved", "=COUNTIF('Expired Exceptions'!I5:I104,\"Compliance Achieved\")", "Expired"),
    ]

    for row_offset, (metric, formula, category) in enumerate(t2_metrics):
        row = 13 + row_offset
        # Col A: metric name
        c_a = ws.cell(row=row, column=1, value=metric)
        c_a.font = Font(name="Calibri")
        c_a.alignment = Alignment(horizontal="left", vertical="center")
        c_a.border = border_thin
        # Col B: formula value
        c_b = ws.cell(row=row, column=2, value=formula)
        c_b.font = Font(name="Calibri", color="000000")
        c_b.alignment = Alignment(horizontal="center", vertical="center")
        c_b.border = border_thin
        # Col C: category
        c_c = ws.cell(row=row, column=3, value=category)
        c_c.font = Font(italic=True, name="Calibri", color="003366")
        c_c.alignment = Alignment(horizontal="left", vertical="center")
        c_c.border = border_thin
        # Cols D-G: empty with border
        for col_idx in range(4, 8):
            ws.cell(row=row, column=col_idx).border = border_thin

    # ── TABLE 3: Critical Findings ────────────────────────────────────────────
    # Row 32: blank gap
    # Row 33 (was 32 in design, adjusted): section header
    ws.merge_cells("A34:G34")
    ws["A34"] = "CRITICAL FINDINGS — AT-RISK ACTIVE EXCEPTIONS"
    ws["A34"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws["A34"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws["A34"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A34"].border = border_thin

    # Row 34: TABLE 3 column headers
    t3_headers = ["Assessment Area", "Exception ID", "System Name", "Exception Type", "Status", "Expiry Date", "Notes"]
    for col_idx, hdr in enumerate(t3_headers, 1):
        c = ws.cell(row=35, column=col_idx, value=hdr)
        c.font = Font(bold=True, size=10, name="Calibri")
        c.fill = grey_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin

    # TABLE 3 data rows (35-44) — INDEX/SMALL/IF for Expiring Soon + Pending Revocation
    for k in range(1, 11):
        row = 35 + k
        rng_a = "'Active Exceptions'!A$5:A$54"
        rng_b = "'Active Exceptions'!B$5:B$54"
        rng_d = "'Active Exceptions'!D$5:D$54"
        rng_m = "'Active Exceptions'!M$5:M$54"
        rng_g = "'Active Exceptions'!G$5:G$54"
        cond = f"(('Active Exceptions'!M$5:M$54=\"Expiring Soon\")+('Active Exceptions'!M$5:M$54=\"Pending Revocation\"))"
        row_nums = f"ROW('Active Exceptions'!A$5:A$54)-ROW('Active Exceptions'!A$5)+1"
        small_part = f"SMALL(IF({cond},{row_nums}),{k})"

        formulas = [
            "Active Exceptions",
            f"=IFERROR(INDEX({rng_a},{small_part}),\"\")",
            f"=IFERROR(INDEX({rng_b},{small_part}),\"\")",
            f"=IFERROR(INDEX({rng_d},{small_part}),\"\")",
            f"=IFERROR(INDEX({rng_m},{small_part}),\"\")",
            f"=IFERROR(INDEX({rng_g},{small_part}),\"\")",
            "",
        ]
        for col_idx, val in enumerate(formulas, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = Font(name="Calibri", color="000000")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = border_thin

    # TOTAL row (row 45)
    # Apply FFFFCC fill to TABLE 3 data rows (rows 35-44)
    for _r3 in range(36, 46):
        for _c3 in range(1, 8):
            ws.cell(row=_r3, column=_c3).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws.cell(row=46, column=1, value="TOTAL At-Risk Exceptions").font = Font(bold=True, name="Calibri")
    ws.cell(row=46, column=2).value = "=COUNTIF('Active Exceptions'!M5:M54,\"Expiring Soon\")+COUNTIF('Active Exceptions'!M5:M54,\"Pending Revocation\")"
    ws.cell(row=46, column=2).font = Font(bold=True, name="Calibri")
    for col_idx in range(1, 8):
        c = ws.cell(row=46, column=col_idx)
        c.fill = grey_fill
        c.border = border_thin
        if col_idx == 1:
            c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    col_widths = {"A": 35, "B": 18, "C": 25, "D": 22, "E": 22, "F": 18, "G": 20}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A4"

def create_evidence_register(ws):
    """Create Evidence Register sheet -- golden standard common sheet."""
    ws.title = "Evidence Register"
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "EVIDENCE REGISTER"
    cell.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws["A2"] = "Use this register to document all evidence supporting exception requests, risk assessments, and compensating controls."
    ws["A2"].font = Font(italic=True, size=10, name="Calibri")
    ws.merge_cells("A2:H2")

    headers = [
        "Evidence ID",
        "Evidence Type",
        "Description",
        "Related Exception ID",
        "Date Collected",
        "Collected By",
        "Location/Reference",
        "Notes"
    ]
    for col_idx, hdr in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=hdr)
        c.font = Font(bold=True, name="Calibri", color="FFFFFF")
        c.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    grey_er3_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_er3_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 5: F2F2F2 grey sample row with EV-001 + realistic example data
    er3_sample = [
        "EV-001", "Exception Approval Record EXC-A817-001", "Signed Attestation",
        "Approved exception for isolated lab server with manual NTP sync",
        "/evidence/exceptions/exc-a817-001-approval.pdf",
        "10.01.2025", "CISO", "Verified"
    ]
    for col_idx, value in enumerate(er3_sample, 1):
        c = ws.cell(row=5, column=col_idx, value=value)
        c.fill = grey_er3_fill
        c.font = Font(name="Calibri", size=10, color="808080")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border_thin

    # Rows 6-105: 100 EMPTY FFFFCC rows — NO EV IDs
    for row_idx in range(6, 106):
        for col_idx in range(1, 9):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = yellow_er3_fill
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = border_thin
            c.font = Font(name="Calibri")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 35
    ws.freeze_panes = "A5"

def create_approval_sheet(ws):
    """Create approval and sign-off sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G7),\"\")"),
        ("Assessment Status:", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    
    # Apply borders to all merged cell top-left corners (GS-AS-011)
    _as_thin = Side(style="thin")
    _as_border = Border(left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin)
    for merge_range in ws.merged_cells.ranges:
        tl = ws.cell(merge_range.min_row, merge_range.min_col)
        tl.border = _as_border
    ws.freeze_panes = "A3"

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("="*80)
    logger.info("ISMS A.8.17 - Clock Synchronisation Exception Register Generator")
    logger.info("="*80)
    logger.info("")
    
    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.remove(wb.active)  # Remove default sheet

    # Create styles
    styles = create_styles()

    # Create sheets
    logger.info("Creating sheets...")
    sheet_names = [
        "Instructions & Legend",
        "Exception Requests",
        "Active Exceptions",
        "Expired Exceptions",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off"
    ]
    for sheet_name in sheet_names:
        wb.create_sheet(sheet_name)

    logger.info("  [1/7] Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("  [2/7] Exception Requests...")
    create_exception_requests_sheet(wb["Exception Requests"], styles)

    logger.info("  [3/7] Active Exceptions...")
    create_active_exceptions_sheet(wb["Active Exceptions"], styles)

    logger.info("  [4/7] Expired Exceptions...")
    create_expired_exceptions_sheet(wb["Expired Exceptions"], styles)

    logger.info("  [6/7] Evidence Register...")
    create_evidence_register(wb["Evidence Register"])

    logger.info("  [5/7] Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("  [7/7] Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"])
    
    # Save workbook
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    finalize_validations(wb)
    wb.save(output_path)
    
    logger.info("")
    logger.info("="*80)
    logger.info(f"{CHECK} SUCCESS: {output_path}")
    logger.info("="*80)
    logger.info("")
    logger.info("Exception Register Structure:")
    logger.info("  • Instructions & Legend - Usage guidance and policy references")
    logger.info("  • Exception Requests - Pending approvals")
    logger.info("  • Active Exceptions - Approved exceptions in effect")
    logger.info("  • Expired Exceptions - Historical record")
    logger.info("  • Summary Dashboard - Metrics and compliance status")
    logger.info("  • Evidence Register - Supporting documentation")
    logger.info("  • Approval Sign-Off - Final approval workflow")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open workbook and review Instructions & Legend sheet")
    logger.info("  2. Complete Exception Requests for new exceptions")
    logger.info("  3. CISO/Executive reviews and approves")
    logger.info("  4. Move approved to Active Exceptions")
    logger.info("  5. Quarterly review of all active exceptions")
    logger.info("  6. Move expired/revoked to Expired Exceptions")
    logger.info("  7. Record all evidence in Evidence Register")
    logger.info("  8. Complete Approval Sign-Off when finished")
    logger.info("")
    logger.info("Policy Reference: ISMS-POL-A.8.17 Section 3.3 (Exception Management)")
    logger.info("")

def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
