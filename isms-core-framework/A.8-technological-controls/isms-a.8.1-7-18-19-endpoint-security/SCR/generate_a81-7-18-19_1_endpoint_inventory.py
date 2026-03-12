#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.1-7-18-19.S1 - Endpoint Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.1, A.8.7, A.8.18, A.8.19
Assessment Domain 1 of 6: Endpoint Device Inventory and Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific endpoint environment, discovery tools, and
inventory requirements.

Key customization areas:
1. Endpoint discovery methods (match your actual tools: MDM, SCCM, Jamf, etc.)
2. Classification categories (adapt to your device types and ownership models)
3. Security baseline requirements (specific to your endpoint standards)
4. Encryption technologies (based on your OS mix and encryption solutions)
5. Management platforms (aligned with your endpoint management infrastructure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.1-7-18-19 Endpoint Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
endpoint device inventory, classification, security baseline compliance, and
management enrollment across all endpoint types.

**Purpose:**
Enables systematic assessment of endpoint inventory completeness and security
posture against ISO 27001:2022 Control A.8.1 requirements, supporting evidence-
based validation of endpoint device management and security controls.

**Assessment Scope:**
- Complete endpoint device inventory (laptops, desktops, mobile, tablets, IoT)
- Device classification (corporate-owned, BYOD, contractor, guest)
- Security baseline compliance per endpoint type
- Encryption status (full disk, file-level, mobile encryption)
- Management platform enrollment (MDM, agent-based, unmanaged)
- Remote wipe capabilities
- Lost/stolen device procedures
- Disposal and decommissioning tracking
- Endpoint access controls (authentication, screen lock)
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and classification framework
2. Inventory - Complete endpoint device inventory with 50 sample rows
3. Classification - Device type, ownership model, criticality analysis
4. Baseline_Compliance - Security baseline compliance per endpoint
5. Encryption_Status - Encryption technology and status tracking
6. Management_Enrollment - MDM/management platform enrollment
7. Capability_Requirements - Endpoint security capability assessment
8. Evidence_Register - Audit evidence tracking and documentation
9. Gap_Analysis - Non-compliant endpoints and remediation requirements
10. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dropdown lists for consistency
- Conditional formatting for compliance status visualization
- Automated gap identification for unmanaged/unencrypted devices
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with endpoint management tool outputs

**Integration:**
This assessment provides the inventory foundation for all other endpoint
security assessments (A.8.7 malware protection, A.8.18 privileged utilities,
A.8.19 software controls) and feeds into the A.8.1-7-18-19.S6 Compliance
Dashboard.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a81-7-18-19_1_endpoint_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a81-7-18-19_1_endpoint_inventory.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a81-7-18-19_1_endpoint_inventory.py --date 20250125

Output:
    File: ISMS-IMP-A.8.1-7-18-19.S1_Endpoint_Inventory_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize classification categories for your organisation
    2. Inventory all endpoint devices using discovery tools
    3. Complete inventory sheet with actual endpoint data
    4. Classify endpoints by type, ownership, and criticality
    5. Assess baseline compliance for each endpoint
    6. Validate encryption status
    7. Verify management platform enrollment
    8. Conduct gap analysis for non-compliant endpoints
    9. Define remediation actions with timelines
    10. Collect and link audit evidence
    11. Obtain stakeholder approvals
    12. Feed results into other assessments and compliance dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.1 (Primary)
Assessment Domain:    1 of 6 (Endpoint Inventory and Classification)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.1-7-18-19: Endpoint Security Framework (Governance)
    - ISMS-IMP-A.8.1-7-18-19-S1: Endpoint Discovery Process
    - ISMS-IMP-A.8.1-7-18-19-S2: Protection Coverage Assessment (Domain 2)
    - ISMS-IMP-A.8.1-7-18-19-S3: Software Controls Assessment (Domain 3)
    - ISMS-IMP-A.8.1-7-18-19-S4: Privileged Utilities Assessment (Domain 4)
    - ISMS-IMP-A.8.1-7-18-19-S5: Compliance Matrix (Domain 5)
    - ISMS-IMP-A.8.1-7-18-19-S6: Compliance Dashboard (Domain 6)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.1-7-18-19-S1 specification
    - Supports comprehensive endpoint inventory and classification
    - Integrated with A.8.1-7-18-19.S6 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Technology Diversity:**
Endpoint environments vary significantly across organisations. This script
provides a generic framework that must be customized for your specific:
- Endpoint types (Windows/Mac/Linux/mobile/IoT mix)
- Management platforms (Intune/Jamf/SCCM/MDM/etc.)
- Encryption solutions (BitLocker/FileVault/LUKS/mobile encryption)
- Ownership models (corporate/BYOD/contractor/guest policies)

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect complete endpoint inventory and classification.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Complete endpoint inventory with serial numbers and users
- Device locations and classifications
- Security posture and vulnerability information
- Organisational structure and access patterns

Handle in accordance with your organisation's data classification policies.

**BYOD Considerations:**
Personal devices require different security controls and privacy protections.
Clearly distinguish BYOD from corporate-owned devices. Ensure BYOD policies
address data separation, limited management scope, and employee privacy rights.

**Maintenance:**
Review and update assessment:
- Monthly: Update inventory for new/decommissioned devices
- Quarterly: Refresh classification and baseline compliance
- Semi-annually: Comprehensive reassessment
- Annually: Full endpoint security review
- Ad-hoc: When significant infrastructure changes occur

**Quality Assurance:**
Have endpoint management team and security engineers validate inventory
completeness before using results for compliance reporting or remediation
decisions. Missing endpoints = missing security coverage.

**Continuous Improvement:**
Use assessment findings to improve endpoint management processes:
- Automate endpoint discovery and enrollment
- Implement continuous compliance monitoring
- Reduce time-to-remediation for gaps
- Improve baseline enforcement mechanisms

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # Green checkmark
XMARK = '\u274C'      # Red X
WARNING = '\u26A0'    # Warning sign
CHART = '[CHART]'    # Chart (BMP-safe)
TARGET = '[TARGET]'  # Target (BMP-safe)
SHIELD = '\u26F2'    # Shield (BMP-safe)
LOCK = '\u26BF'      # Lock (BMP-safe)
LAPTOP = '[LAPTOP]'  # Laptop (BMP-safe)
VIRUS = '[VIRUS]'    # Virus (BMP-safe)
BULLET = '\u2022'     # Bullet point
ARROW = '\u2192'      # Right arrow
# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.1-7-18-19.1"
WORKBOOK_NAME = "Endpoint Inventory"
CONTROL_ID   = "A.8.1-7-18-19"
CONTROL_NAME = "Endpoint and Device Security"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure for Endpoint Inventory Assessment
    sheets = [
        "Instructions & Legend",
        "Inventory",
        "Baseline Compliance",
        "Encryption Status",
        "Management Enrollment",
        "Capability Requirements",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_unknown": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "gap_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "gap_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "gap_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass


def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_unknown': DataValidation(
            type="list",
            formula1='"Yes,No,Unknown"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1='"Compliant,Partial,Non-Compliant,Unknown"',
            allow_blank=False
        ),
        'device_type': DataValidation(
            type="list",
            formula1='"Laptop,Desktop,Mobile (iOS),Mobile (Android),Tablet,Workstation,Other"',
            allow_blank=False
        ),
        'operating_system': DataValidation(
            type="list",
            formula1='"Windows 11,Windows 10,macOS 14 (Sonoma),macOS 13 (Ventura),iOS 17,iOS 16,Android 14,Android 13,Linux (Ubuntu),Linux (RHEL/CentOS),Other"',
            allow_blank=False
        ),
        'ownership_model': DataValidation(
            type="list",
            formula1='"Corporate-Owned,BYOD,Contractor,Shared Device,Other"',
            allow_blank=False
        ),
        'criticality': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'encryption_status': DataValidation(
            type="list",
            formula1='"Encrypted,Not Encrypted,Partial,Unknown"',
            allow_blank=False
        ),
        'encryption_technology': DataValidation(
            type="list",
            formula1='"BitLocker,FileVault,LUKS,Built-in (iOS/Android),Third-Party,None,Unknown"',
            allow_blank=False
        ),
        'mdm_platform': DataValidation(
            type="list",
            formula1='"Microsoft Intune,Jamf Pro,SCCM,VMware Workspace ONE,MobileIron,Citrix Endpoint Management,Kandji,Google Workspace,None,Other"',
            allow_blank=False
        ),
        'enrollment_status': DataValidation(
            type="list",
            formula1='"Enrolled,Not Enrolled,Partial,Pending,Unknown"',
            allow_blank=False
        ),
        'baseline_compliance': DataValidation(
            type="list",
            formula1='"Compliant,Partial,Non-Compliant,Remediation in Progress,Unknown"',
            allow_blank=False
        ),
        'firewall_status': DataValidation(
            type="list",
            formula1='"Enabled,Disabled,Partial,Unknown"',
            allow_blank=False
        ),
        'antivirus_status': DataValidation(
            type="list",
            formula1='"Active,Outdated,Not Installed,Inactive,Unknown"',
            allow_blank=False
        ),
        'device_location': DataValidation(
            type="list",
            formula1='"Office,Remote,Mobile,International,Unknown"',
            allow_blank=False
        ),
        'device_status': DataValidation(
            type="list",
            formula1='"Active,Inactive,Retired,Storage,Pending Setup,Lost/Stolen,Unknown"',
            allow_blank=False
        ),
        'gap_severity': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Closed,On Hold"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Config Export,Screenshot,Report,Certificate,Policy,Log File,Encryption Key Backup,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1='"Verified,Pending,Not Verified,Needs Review"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Inventory Sheet — Complete endpoint inventory from MDM/SCCM/Jamf exports. Populate Device ID, Hostname, Type, OS, User, Location, Status for ALL endpoints.',
        '2. Baseline Compliance — Assess each endpoint against security baseline (OS hardening, firewall, encryption, antivirus). Use compliance status dropdowns.',
        '3. Encryption Status — Document encryption technology per endpoint (BitLocker, FileVault, LUKS, built-in). Verify encryption is active.',
        '4. Management Enrollment — Document MDM platform enrollment status. Identify unmanaged devices.',
        '5. Capability Requirements — Map policy requirements to implementation. Verify all requirements are addressed.',
        '6. Evidence Register — Document ALL evidence (config exports, screenshots, reports, certificates). Minimum 20 evidence entries.',
        '7. Gap Analysis — Document identified gaps with severity, root cause, and remediation plan. Track remediation progress.',
        '8. Approval Sign-Off — Obtain approvals: Assessor → IT Operations Manager → CISO. Final approval required for assessment closure.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A22"] = "Status Legend"
    ws["A22"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=23, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 24 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_inventory_sheet(ws, styles):
    """Create main endpoint inventory sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = "ENDPOINT DEVICE GROUP INVENTORY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:N2')
    cell = ws['A2']
    cell.value = "Group-based inventory — each row represents a device group (AD OU, Entra ID group, department group, or device type)"
    apply_style(cell, styles['subheader'])

    # Row 3: insert-row guidance note
    ws.merge_cells('A3:N3')
    cell = ws['A3']
    cell.value = "NOTE: To add more device groups, insert rows within the yellow range (rows 6–55) — do not insert above row 6 or below row 55 to preserve formula integrity."
    cell.font = Font(name="Calibri", size=9, italic=True, color="003366")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 20

    # Column headers
    headers = [
        "Group ID",
        "Device Group Name",
        "Group Type",
        "Device Type",
        "OS Platform",
        "Device Count",
        "Department",
        "Ownership Model",
        "Location",
        "Management Platform",
        "Compliance Status",
        "Last Assessed",
        "Criticality",
        "Notes",
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 5: F2F2F2 grey sample row (Golden Standard — 1 row, realistic data)
    sample_row = 5
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_s = Side(style="thin")
    border_s = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    sample_data = [
        "GRP-001", "SG-Windows-Endpoints",
        "Entra ID Group", "Workstations", "Windows", 312,
        "IT", "Corporate-Owned", "Office",
        "Microsoft Intune", f"{CHECK} Compliant", "20.02.2026",
        "High", "Production Windows workstations — Intune managed. See evidence ER-001.",
    ]
    for col, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=sample_row, column=col)
        cell.value = val
        cell.fill = grey_fill
        cell.border = border_s
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Inline DVs for new group-based columns
    group_type_dv = DataValidation(
        type="list",
        formula1='"AD OU,Entra ID Group,Department Group,Location Group,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(group_type_dv)

    os_platform_dv = DataValidation(
        type="list",
        formula1='"Windows,macOS,Linux,iOS,Android,Mixed,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(os_platform_dv)

    inv_compliance_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,Unknown"',
        allow_blank=True,
    )
    ws.add_data_validation(inv_compliance_dv)

    # Rows 6-55: 50 empty FFFFCC yellow input rows
    start_row = 6
    for i in range(50):
        current_row = start_row + i

        # Group ID (manual input — no auto-fill)
        cell = ws.cell(row=current_row, column=1)
        apply_style(cell, styles['input_cell'])

        # Device Group Name (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])

        # Group Type (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        group_type_dv.add(cell)

        # Device Type (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['device_type'].add(cell)

        # OS Platform (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        os_platform_dv.add(cell)

        # Device Count (input — numeric)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])

        # Department (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])

        # Ownership Model (dropdown)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        validations['ownership_model'].add(cell)

        # Location (dropdown)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        validations['device_location'].add(cell)

        # Management Platform (dropdown)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        validations['mdm_platform'].add(cell)

        # Compliance Status (dropdown)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])
        inv_compliance_dv.add(cell)

        # Last Assessed (input)
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])

        # Criticality (dropdown)
        cell = ws.cell(row=current_row, column=13)
        apply_style(cell, styles['input_cell'])
        validations['criticality'].add(cell)

        # Notes (input)
        cell = ws.cell(row=current_row, column=14)
        apply_style(cell, styles['input_cell'])

    # Summary statistics (COUNTIF ranges start at row 6 to exclude sample row)
    summary_row = 58
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} INVENTORY SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Device Groups:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = '=COUNTA(B6:B55)'

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Compliant Groups:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"{CHECK} Compliant")'

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{WARNING} Partial Compliance:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"{WARNING} Partial")'

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Non-Compliant Groups:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"{XMARK} Non-Compliant")'

    # Column widths (14 cols A–N)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 28
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 35

    # Freeze panes
    ws.freeze_panes = 'A5'


# ============================================================================
# SECTION 5: REMOVED — Classification sheet removed; key metrics moved to
#            Summary Dashboard TABLE 2 (Corporate-Owned / BYOD / High-Criticality KPIs)
# ============================================================================


# ============================================================================
# SECTION 6: BASELINE COMPLIANCE SHEET
# ============================================================================

def create_baseline_compliance_sheet(ws, styles):
    """Create per-endpoint baseline compliance assessment sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "ENDPOINT BASELINE COMPLIANCE ASSESSMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "Per-endpoint security baseline compliance (OS hardening, firewall, antivirus, configuration)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Group ID",
        "Device Group Name",
        "Device Type",
        "OS Hardening",
        "Firewall Status",
        "Antivirus / EDR",
        "Screen Lock",
        "Patch Compliance",
        "Password Policy",
        "Baseline Profile",
        "Compliance Status",
        "Notes",
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 5: F2F2F2 grey sample row
    sample_row = 5
    grey_fill_bc = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_bc = Side(style="thin")
    border_bc = Border(left=thin_bc, right=thin_bc, top=thin_bc, bottom=thin_bc)
    sample_bc_data = [
        "GRP-001", "SG-Windows-Endpoints", "Workstations",
        f"{CHECK} Compliant", f"{CHECK} Enabled", f"{CHECK} Active",
        "Yes", f"{CHECK} Current", f"{CHECK} Compliant",
        "CIS Level 1", f"{CHECK} Compliant",
        "Example group row — CIS L1 baseline applied via Intune. See ER-002.",
    ]
    for col, val in enumerate(sample_bc_data, start=1):
        cell = ws.cell(row=sample_row, column=col)
        cell.value = val
        cell.fill = grey_fill_bc
        cell.border = border_bc
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Inline DVs for emoji-based status fields
    bc_status_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(bc_status_dv)

    bc_firewall_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Enabled,{WARNING} Partial,{XMARK} Disabled,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(bc_firewall_dv)

    bc_av_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Active,{WARNING} Outdated,{XMARK} Not Installed,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(bc_av_dv)

    bc_patch_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Current,{WARNING} Partial,{XMARK} Outdated,Unknown"',
        allow_blank=True,
    )
    ws.add_data_validation(bc_patch_dv)

    bc_screen_dv = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(bc_screen_dv)

    # Rows 6-55: 50 empty FFFFCC rows, linked to Inventory
    start_row = 6
    for i in range(50):
        current_row = start_row + i
        row_inv = 6 + i

        # Group ID (linked to Inventory — IF wrapper prevents ghost zeros)
        ws.cell(row=current_row, column=1).value = f'=IF(Inventory!A{row_inv}="","",Inventory!A{row_inv})'

        # Device Group Name (linked to Inventory)
        ws.cell(row=current_row, column=2).value = f'=IF(Inventory!B{row_inv}="","",Inventory!B{row_inv})'

        # Device Type (linked to Inventory col D — new position after redesign)
        ws.cell(row=current_row, column=3).value = f'=IF(Inventory!D{row_inv}="","",Inventory!D{row_inv})'

        # Apply FFFFCC fill + border + alignment to linked cols A–C
        yfill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(1, 4):
            _c = ws.cell(row=current_row, column=c)
            _c.fill = yfill
            _c.border = border_bc
            _c.alignment = Alignment(horizontal="left", vertical="center")

        # OS Hardening (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        bc_status_dv.add(cell)

        # Firewall Status (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        bc_firewall_dv.add(cell)

        # Antivirus / EDR (dropdown)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        bc_av_dv.add(cell)

        # Screen Lock (dropdown)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        bc_screen_dv.add(cell)

        # Patch Compliance (dropdown)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        bc_patch_dv.add(cell)

        # Password Policy (dropdown)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        bc_status_dv.add(cell)

        # Baseline Profile (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])

        # Compliance Status (calculated — wildcards on emoji prefix)
        cell = ws.cell(row=current_row, column=11)
        cell.value = (
            f'=IF(COUNTA(D{current_row}:I{current_row})=0,"",'
            f'IF(COUNTIF(D{current_row}:I{current_row},"{XMARK}*")+COUNTIF(G{current_row}:G{current_row},"No")>0,'
            f'"{XMARK} Non-Compliant",'
            f'IF(COUNTIF(D{current_row}:I{current_row},"{WARNING}*")>0,'
            f'"{WARNING} Partial","{CHECK} Compliant")))'
        )

        # Notes (input)
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])

    # Summary statistics (COUNTIF ranges start at row 6 to exclude sample row)
    summary_row = 58
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} BASELINE COMPLIANCE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Compliant Groups:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"{CHECK} Compliant")'

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{WARNING} Partial Compliance:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"{WARNING} Partial")'

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Non-Compliant Groups:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"{XMARK} Non-Compliant")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "Compliance Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF((COUNTIF(K6:K55,"{CHECK} Compliant")+COUNTIF(K6:K55,"{WARNING} Partial")+COUNTIF(K6:K55,"{XMARK} Non-Compliant"))=0,0,COUNTIF(K6:K55,"{CHECK} Compliant")/(COUNTIF(K6:K55,"{CHECK} Compliant")+COUNTIF(K6:K55,"{WARNING} Partial")+COUNTIF(K6:K55,"{XMARK} Non-Compliant")))&"%"'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 30

    # Freeze panes (gold standard: data rows start at row 5, freeze A5)
    ws.freeze_panes = 'A5'


# ============================================================================
# SECTION 7: ENCRYPTION STATUS SHEET
# ============================================================================

def create_encryption_status_sheet(ws, styles):
    """Create encryption status and key management sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "ENDPOINT ENCRYPTION STATUS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Encryption technology, status, key management, and recovery procedures per endpoint"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Device Type",
        "Encryption Technology",
        "Encryption Status",
        "Full Disk Encryption",
        "Key Escrow/Recovery",
        "Recovery Key Location",
        "Last Verified Date",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 5: F2F2F2 grey sample row with realistic example data
    sample_row_enc = 5
    grey_fill_enc = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_enc = Side(style="thin")
    border_enc = Border(left=thin_enc, right=thin_enc, top=thin_enc, bottom=thin_enc)
    sample_enc_data = [
        "GRP-001", "SG-Windows-Endpoints", "Workstations",
        "BitLocker", f"{CHECK} Encrypted", "Yes",
        "Microsoft Intune", "Azure AD / Intune portal", "20.02.2026",
        "Enforced via Intune policy — BitLocker recovery keys in Entra ID.",
    ]
    for col, val in enumerate(sample_enc_data, start=1):
        cell = ws.cell(row=sample_row_enc, column=col)
        cell.value = val
        cell.fill = grey_fill_enc
        cell.border = border_enc
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Inline DVs for encryption status fields
    enc_status_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Encrypted,{WARNING} Partial,{XMARK} Not Encrypted,Unknown"',
        allow_blank=True,
    )
    ws.add_data_validation(enc_status_dv)

    enc_fde_dv = DataValidation(
        type="list",
        formula1='"Yes,No,Partial,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(enc_fde_dv)

    enc_escrow_dv = DataValidation(
        type="list",
        formula1='"Microsoft Intune,Azure AD,On-premise AD,HashiCorp Vault,Manual,None"',
        allow_blank=True,
    )
    ws.add_data_validation(enc_escrow_dv)

    # Rows 6-55: 50 empty FFFFCC rows, linked to Inventory
    start_row = 6
    for i in range(50):
        current_row = start_row + i
        row_inv = 6 + i

        # Group ID (linked to Inventory — IF wrapper prevents ghost zeros)
        ws.cell(row=current_row, column=1).value = f'=IF(Inventory!A{row_inv}="","",Inventory!A{row_inv})'

        # Device Group Name (linked to Inventory)
        ws.cell(row=current_row, column=2).value = f'=IF(Inventory!B{row_inv}="","",Inventory!B{row_inv})'

        # Device Type (linked to Inventory col D — new position after redesign)
        ws.cell(row=current_row, column=3).value = f'=IF(Inventory!D{row_inv}="","",Inventory!D{row_inv})'

        # Apply FFFFCC fill + border + alignment to linked cols A–C
        yfill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(1, 4):
            _c = ws.cell(row=current_row, column=c)
            _c.fill = yfill
            _c.border = border_enc
            _c.alignment = Alignment(horizontal="left", vertical="center")

        # Encryption Technology (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['encryption_technology'].add(cell)

        # Encryption Status (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        enc_status_dv.add(cell)

        # Full Disk Encryption (dropdown)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        enc_fde_dv.add(cell)

        # Key Escrow / Recovery Location (dropdown)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        enc_escrow_dv.add(cell)

        # Recovery Key Location (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])

        # Last Verified Date (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])

        # Notes (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])

    # Summary statistics (COUNTIF ranges start at row 6 to exclude sample row)
    summary_row = 58
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} ENCRYPTION COVERAGE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Encrypted Groups:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"{CHECK} Encrypted")'

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Not Encrypted:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"{XMARK} Not Encrypted")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "Encryption Coverage Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF((COUNTIF(E6:E55,"{CHECK} Encrypted")+COUNTIF(E6:E55,"{WARNING} Partial")+COUNTIF(E6:E55,"{XMARK} Not Encrypted"))=0,0,COUNTIF(E6:E55,"{CHECK} Encrypted")/(COUNTIF(E6:E55,"{CHECK} Encrypted")+COUNTIF(E6:E55,"{WARNING} Partial")+COUNTIF(E6:E55,"{XMARK} Not Encrypted")))&"%"'

    summary_row += 2
    ws[f'A{summary_row}'].value = "BitLocker (Windows):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"BitLocker")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "FileVault (macOS):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"FileVault")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "Built-in (iOS/Android):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"Built-in (iOS/Android)")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "LUKS (Linux):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"LUKS")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 30

    # Freeze panes (gold standard: data rows start at row 5, freeze A5)
    ws.freeze_panes = 'A5'


# ============================================================================
# SECTION 8: MANAGEMENT ENROLLMENT SHEET
# ============================================================================

def create_management_enrollment_sheet(ws, styles):
    """Create MDM/endpoint management enrollment sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "ENDPOINT MANAGEMENT ENROLLMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "MDM platform enrollment status, capabilities, and management coverage"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Group ID",
        "Device Group Name",
        "Device Type",
        "MDM Platform",
        "Enrollment Status",
        "Remote Wipe Capable",
        "Policy Enforcement",
        "Compliance Policy Applied",
        "Last Check-In",
        "MDM Compliance Status",
        "Notes",
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 5: F2F2F2 grey sample row with realistic example data
    sample_row_me = 5
    grey_fill_me = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_me = Side(style="thin")
    border_me = Border(left=thin_me, right=thin_me, top=thin_me, bottom=thin_me)
    sample_me_data = [
        "GRP-001", "SG-Windows-Endpoints", "Workstations",
        "Microsoft Intune", f"{CHECK} Enrolled", "Yes",
        "Yes", "Windows Compliance Policy v3", "20.02.2026",
        f"{CHECK} Compliant", "All enrolled via Autopilot — see evidence ER-003.",
    ]
    for col, val in enumerate(sample_me_data, start=1):
        cell = ws.cell(row=sample_row_me, column=col)
        cell.value = val
        cell.fill = grey_fill_me
        cell.border = border_me
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Inline DVs for management enrollment status fields
    me_enroll_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Enrolled,{WARNING} Partial,{XMARK} Not Enrolled,Pending,Unknown"',
        allow_blank=True,
    )
    ws.add_data_validation(me_enroll_dv)

    me_policy_dv = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=True,
    )
    ws.add_data_validation(me_policy_dv)

    me_mdm_compliance_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{XMARK} Non-Compliant,Unknown"',
        allow_blank=True,
    )
    ws.add_data_validation(me_mdm_compliance_dv)

    # Rows 6-55: 50 empty FFFFCC rows, linked to Inventory
    start_row = 6
    for i in range(50):
        current_row = start_row + i
        row_inv = 6 + i

        # Group ID (linked to Inventory — IF wrapper prevents ghost zeros)
        ws.cell(row=current_row, column=1).value = f'=IF(Inventory!A{row_inv}="","",Inventory!A{row_inv})'

        # Device Group Name (linked to Inventory)
        ws.cell(row=current_row, column=2).value = f'=IF(Inventory!B{row_inv}="","",Inventory!B{row_inv})'

        # Device Type (linked to Inventory col D — new position after redesign)
        ws.cell(row=current_row, column=3).value = f'=IF(Inventory!D{row_inv}="","",Inventory!D{row_inv})'

        # Apply FFFFCC fill + border + alignment to linked cols A–C
        yfill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(1, 4):
            _c = ws.cell(row=current_row, column=c)
            _c.fill = yfill
            _c.border = border_me
            _c.alignment = Alignment(horizontal="left", vertical="center")

        # MDM Platform (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['mdm_platform'].add(cell)

        # Enrollment Status (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        me_enroll_dv.add(cell)

        # Remote Wipe Capable (dropdown)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)

        # Policy Enforcement (dropdown)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        me_policy_dv.add(cell)

        # Compliance Policy Applied (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])

        # Last Check-In (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])

        # MDM Compliance Status (dropdown)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        me_mdm_compliance_dv.add(cell)

        # Notes (input)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])

    # Summary statistics (COUNTIF ranges start at row 6 to exclude sample row)
    summary_row = 58
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} MANAGEMENT ENROLLMENT SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Enrolled Groups:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"{CHECK} Enrolled")'

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Not Enrolled:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"{XMARK} Not Enrolled")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "Enrollment Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF((COUNTIF(E6:E55,"{CHECK} Enrolled")+COUNTIF(E6:E55,"{WARNING} Partial")+COUNTIF(E6:E55,"{XMARK} Not Enrolled"))=0,0,COUNTIF(E6:E55,"{CHECK} Enrolled")/(COUNTIF(E6:E55,"{CHECK} Enrolled")+COUNTIF(E6:E55,"{WARNING} Partial")+COUNTIF(E6:E55,"{XMARK} Not Enrolled")))&"%"'

    summary_row += 2
    ws[f'A{summary_row}'].value = "By MDM Platform:"
    ws[f'A{summary_row}'].font = Font(bold=True)

    summary_row += 1
    ws[f'A{summary_row}'].value = "Microsoft Intune:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"Microsoft Intune")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "Jamf Pro:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"Jamf Pro")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "SCCM:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"SCCM")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "None:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D6:D55,"None")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 30

    # Freeze panes (gold standard: data rows start at row 5, freeze A5)
    ws.freeze_panes = 'A5'


# ============================================================================
# SECTION 9: CAPABILITY REQUIREMENTS SHEET
# ============================================================================

def create_capability_requirements_sheet(ws, styles):
    """Create policy requirements mapping sheet (A.8.1 requirements)."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPABILITY REQUIREMENTS MAPPING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "A.8.1 Policy Requirements → Implementation Verification (30 requirements)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Req ID",
        "Policy Requirement",
        "Implemented",
        "Evidence Reference",
        "Notes",
        "Status"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Requirements (from POL-S2)
    requirements = [
        ("REQ-A81-001", "Complete endpoint inventory maintained"),
        ("REQ-A81-002", "Endpoints classified by type, ownership, criticality"),
        ("REQ-A81-003", "Security baselines defined per endpoint type"),
        ("REQ-A81-004", "Baseline compliance ≥95% for corporate endpoints"),
        ("REQ-A81-005", "Full disk encryption for laptops (≥95%)"),
        ("REQ-A81-006", "Full disk encryption for desktops (≥90%)"),
        ("REQ-A81-007", "Mobile devices encrypted (≥95%)"),
        ("REQ-A81-008", "Encryption key escrow/recovery implemented"),
        ("REQ-A81-009", "MDM enrollment ≥95% corporate endpoints"),
        ("REQ-A81-010", "MDM enrollment ≥80% BYOD (if BYOD allowed)"),
        ("REQ-A81-011", "Remote wipe capability for all corporate mobile devices"),
        ("REQ-A81-012", "Screen lock enforced (max 15 minutes)"),
        ("REQ-A81-013", "Strong password policy enforced"),
        ("REQ-A81-014", "Firewall enabled on all endpoints"),
        ("REQ-A81-015", "Anti-malware deployed (see A.8.7)"),
        ("REQ-A81-016", "Automatic security updates enabled"),
        ("REQ-A81-017", "Quarterly endpoint inventory reconciliation"),
        ("REQ-A81-018", "Lost/stolen device procedure documented"),
        ("REQ-A81-019", "Lost/stolen incidents responded to within 24 hours"),
        ("REQ-A81-020", "Remote wipe initiated within 24 hours (if device not recovered)"),
        ("REQ-A81-021", "Secure disposal procedure for all endpoints"),
        ("REQ-A81-022", "Disposal certificates retained"),
        ("REQ-A81-023", "BYOD user agreements signed (if BYOD allowed)"),
        ("REQ-A81-024", "BYOD devices containerized (corporate data separated)"),
        ("REQ-A81-025", "Unmanaged endpoints identified quarterly"),
        ("REQ-A81-026", "Endpoint lifecycle management process"),
        ("REQ-A81-027", "Endpoint procurement includes security requirements"),
        ("REQ-A81-028", "Endpoints configured before user deployment"),
        ("REQ-A81-029", "Endpoint decommissioning includes data sanitization"),
        ("REQ-A81-030", "Annual endpoint security assessment"),
    ]

    # Row 5: F2F2F2 grey sample row
    sample_row_cr = 5
    grey_fill_cr = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_cr = Side(style="thin")
    border_cr = Border(left=thin_cr, right=thin_cr, top=thin_cr, bottom=thin_cr)
    sample_cr_data = ["REQ-A81-000", "Complete endpoint inventory maintained (example sample row)", "Yes", "/evidence/A.8.1/example.png", "Sample row — do not edit", ""]
    for col, val in enumerate(sample_cr_data, start=1):
        cell = ws.cell(row=sample_row_cr, column=col)
        cell.value = val
        cell.fill = grey_fill_cr
        cell.border = border_cr
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # Status formula for sample row
    ws.cell(row=sample_row_cr, column=6).value = f'=IF(C{sample_row_cr}="Yes","{CHECK} Compliant",IF(C{sample_row_cr}="N/A","N/A","{XMARK} Gap"))'

    # Rows 6-35: 30 pre-populated requirement rows
    start_row = 6
    for i, (req_id, requirement) in enumerate(requirements):
        current_row = start_row + i

        # Req ID
        ws.cell(row=current_row, column=1).value = req_id

        # Policy Requirement
        ws.cell(row=current_row, column=2).value = requirement
        ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)

        # Implemented (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)

        # Evidence Reference (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])

        # Notes (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])

        # Status (calculated)
        cell = ws.cell(row=current_row, column=6)
        cell.value = f'=IF(C{current_row}="Yes","{CHECK} Compliant",IF(C{current_row}="N/A","N/A","{XMARK} Gap"))'

        thin_r = Side(style="thin")
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = Border(left=thin_r, right=thin_r, top=thin_r, bottom=thin_r)

    # Summary (COUNTIF ranges start at row 6 to exclude sample row)
    # Requirement rows are 6-35 (30 requirements)
    summary_row = start_row + len(requirements) + 2  # row 38
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} REQUIREMENTS COMPLIANCE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Requirements:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = len(requirements)

    summary_row += 1
    ws[f'A{summary_row}'].value = "Implemented:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(C6:C{start_row+len(requirements)-1},"Yes")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "Not Implemented:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(C6:C{start_row+len(requirements)-1},"No")'

    summary_row += 1
    ws[f'A{summary_row}'].value = "Compliance Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(COUNT(B{summary_row-2}:B{summary_row-1})>0,B{summary_row-2}/B{summary_row-3}*100,0)&"%"'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15

    # Freeze panes (gold standard: data rows start at row 5, freeze A5)
    ws.freeze_panes = 'A5'


# ============================================================================
# SECTION 10: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Gold Standard Summary Dashboard for A.8.1 Endpoint Inventory."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # -------------------------------------------------------------------------
    # Row 1: Title header
    # -------------------------------------------------------------------------
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "ENDPOINT INVENTORY \u2014 SUMMARY DASHBOARD"
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border_thin
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle
    ws.merge_cells('A2:G2')
    cell = ws['A2']
    cell.value = "Consolidated compliance metrics for endpoint device security assessment"
    cell.font = Font(name="Calibri", size=10, italic=True, color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------------------
    # TABLE 1 — COMPLIANCE ASSESSMENT SUMMARY
    # -------------------------------------------------------------------------
    # Row 3: TABLE 1 banner
    ws.merge_cells('A3:G3')
    cell = ws['A3']
    cell.value = "TABLE 1 — COMPLIANCE ASSESSMENT SUMMARY"
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.border = border_thin

    # Row 4: column headers
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, hdr in enumerate(t1_headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Rows 5-8: TABLE 1 data rows (no fill, blue text, thin border)
    t1_data = [
        (
            "Baseline Compliance",
            "=SUM(C5:F5)",
            f"=COUNTIF('Baseline Compliance'!K6:K55,\"{CHECK} Compliant\")",
            f"=COUNTIF('Baseline Compliance'!K6:K55,\"{WARNING} Partial\")",
            f"=COUNTIF('Baseline Compliance'!K6:K55,\"{XMARK} Non-Compliant\")",
            f"=COUNTIF('Baseline Compliance'!K6:K55,\"N/A\")",
            "=IF((B5-F5)=0,0,C5/(B5-F5))",
        ),
        (
            "Encryption (Full-Disk)",
            "=SUM(C6:F6)",
            f"=COUNTIF('Encryption Status'!E6:E55,\"{CHECK} Encrypted\")",
            f"=COUNTIF('Encryption Status'!E6:E55,\"{WARNING} Partial\")",
            f"=COUNTIF('Encryption Status'!E6:E55,\"{XMARK} Not Encrypted\")",
            "0",
            "=IF((B6-F6)=0,0,C6/(B6-F6))",
        ),
        (
            "MDM / UEM Enrollment",
            "=SUM(C7:F7)",
            f"=COUNTIF('Management Enrollment'!E6:E55,\"{CHECK} Enrolled\")",
            f"=COUNTIF('Management Enrollment'!E6:E55,\"{WARNING} Partial\")",
            f"=COUNTIF('Management Enrollment'!E6:E55,\"{XMARK} Not Enrolled\")",
            "0",
            "=IF((B7-F7)=0,0,C7/(B7-F7))",
        ),
        (
            "Capability Requirements",
            "=SUM(C8:F8)",
            f"=COUNTIF('Capability Requirements'!F6:F35,\"{CHECK} Compliant\")",
            "0",
            f"=COUNTIF('Capability Requirements'!F6:F35,\"{XMARK} Gap\")",
            f"=COUNTIF('Capability Requirements'!F6:F35,\"N/A\")",
            "=IF((B8-F8)=0,0,C8/(B8-F8))",
        ),
    ]
    for row_idx, row_data in enumerate(t1_data, start=5):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Calibri", size=10, color="000000")
            cell.fill = PatternFill(fill_type=None)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_idx == 7:
                cell.number_format = "0.0%"

    # Row 9: TOTAL row
    total_vals = ["TOTAL", "=SUM(B5:B8)", "=SUM(C5:C8)", "=SUM(D5:D8)", "=SUM(E5:E8)", "=SUM(F5:F8)", "=IF((B9-F9)=0,0,C9/(B9-F9))"]
    for col_idx, val in enumerate(total_vals, start=1):
        cell = ws.cell(row=9, column=col_idx)
        cell.value = val
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_idx == 7:
            cell.number_format = "0.0%"

    # Row 10: empty gap
    for col_idx in range(1, 8):
        ws.cell(row=10, column=col_idx).value = None

    # -------------------------------------------------------------------------
    # TABLE 2 — KEY PERFORMANCE INDICATORS
    # -------------------------------------------------------------------------
    # Row 11: TABLE 2 banner
    ws.merge_cells('A11:G11')
    cell = ws['A11']
    cell.value = "TABLE 2 — KEY PERFORMANCE INDICATORS"
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.border = border_thin

    # Row 12: column headers
    t2_headers = ["KPI", "Current Value", "Target", "Status", "Last Updated", "Owner", "Notes"]
    for col, hdr in enumerate(t2_headers, start=1):
        cell = ws.cell(row=12, column=col)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # KPI data rows 13-21 (formula in col B, target in col C, rest FFFFCC empty)
    # NOTE: Rows 14 and 19 use COUNTIF-based denominators to avoid FML-SD-001
    #   (Baseline Compliance col K and Capability Requirements col F are auto-formula cols)
    kpi_rows = [
        (13, "Total Endpoints",
         "=COUNTA('Inventory'!B6:B55)",
         ">=50", False),
        (14, "Baseline Compliance Rate",
         f"=IF((COUNTIF('Baseline Compliance'!K6:K55,\"{CHECK} Compliant\")+COUNTIF('Baseline Compliance'!K6:K55,\"{WARNING} Partial\")+COUNTIF('Baseline Compliance'!K6:K55,\"{XMARK} Non-Compliant\"))=0,0,"
         f"COUNTIF('Baseline Compliance'!K6:K55,\"{CHECK} Compliant\")/(COUNTIF('Baseline Compliance'!K6:K55,\"{CHECK} Compliant\")+COUNTIF('Baseline Compliance'!K6:K55,\"{WARNING} Partial\")+COUNTIF('Baseline Compliance'!K6:K55,\"{XMARK} Non-Compliant\")))",
         ">=95%", True),
        (15, "Encryption Coverage",
         f"=IF(COUNTA('Encryption Status'!E6:E55)=0,0,COUNTIF('Encryption Status'!E6:E55,\"{CHECK} Encrypted\")/COUNTA('Encryption Status'!E6:E55))",
         "100%", True),
        (16, "MDM/UEM Enrollment Rate",
         f"=IF(COUNTA('Management Enrollment'!E6:E55)=0,0,COUNTIF('Management Enrollment'!E6:E55,\"{CHECK} Enrolled\")/COUNTA('Management Enrollment'!E6:E55))",
         ">=95%", True),
        (17, "Critical Endpoints",
         "=COUNTIF('Inventory'!M6:M55,\"Critical\")",
         "\u2014", False),
        (18, "Unencrypted Critical Endpoints",
         f"=COUNTIFS('Inventory'!M6:M55,\"Critical\",'Encryption Status'!E6:E55,\"{XMARK} Not Encrypted\")",
         "0", False),
        (19, "Policy Requirements Implemented",
         f"=IF((COUNTIF('Capability Requirements'!C6:C35,\"Yes\")+COUNTIF('Capability Requirements'!C6:C35,\"No\")+COUNTIF('Capability Requirements'!C6:C35,\"N/A\"))=0,0,"
         f"COUNTIF('Capability Requirements'!C6:C35,\"Yes\")/(COUNTIF('Capability Requirements'!C6:C35,\"Yes\")+COUNTIF('Capability Requirements'!C6:C35,\"No\")+COUNTIF('Capability Requirements'!C6:C35,\"N/A\")))",
         "100%", True),
        (20, "Open Gaps",
         "=COUNTIF('Gap Analysis'!K6:K55,\"Open\")+COUNTIF('Gap Analysis'!K6:K55,\"In Progress\")",
         "0", False),
        (21, "Critical/High Gaps",
         "=COUNTIF('Gap Analysis'!E6:E55,\"Critical\")+COUNTIF('Gap Analysis'!E6:E55,\"High\")",
         "0", False),
        # Rows 22-24: Ownership/criticality KPIs (formerly Classification sheet)
        (22, "Corporate-Owned Device Groups",
         "=COUNTIF('Inventory'!H6:H55,\"Corporate-Owned\")",
         "\u2014", False),
        (23, "BYOD Device Groups",
         "=COUNTIF('Inventory'!H6:H55,\"BYOD\")",
         "0 preferred", False),
        (24, "High-Criticality Device Groups",
         "=COUNTIF('Inventory'!M6:M55,\"High\")",
         "\u2014", False),
    ]
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for row_num, kpi_name, formula, target, is_pct in kpi_rows:
        # Col A: KPI name
        cell = ws.cell(row=row_num, column=1)
        cell.value = kpi_name
        cell.font = Font(name="Calibri", size=10, color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center")
        # Col B: formula
        cell = ws.cell(row=row_num, column=2)
        cell.value = formula
        cell.font = Font(name="Calibri", size=10, color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_pct:
            cell.number_format = "0.0%"
        # Col C: target
        cell = ws.cell(row=row_num, column=3)
        cell.value = target
        cell.font = Font(name="Calibri", size=10, color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Cols D-G: FFFFCC empty input cells
        for col_idx in range(4, 8):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = yellow_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Rows 25-26: buffer rows — cols A/B/C white (match data rows), D-G FFFFCC
    for buf_row in range(25, 27):
        for col_idx in range(1, 4):
            cell = ws.cell(row=buf_row, column=col_idx)
            cell.border = border_thin
        for col_idx in range(4, 8):
            cell = ws.cell(row=buf_row, column=col_idx)
            cell.fill = yellow_fill
            cell.border = border_thin

    # Row 27: empty gap
    for col_idx in range(1, 8):
        ws.cell(row=27, column=col_idx).value = None

    # -------------------------------------------------------------------------
    # TABLE 3 — CRITICAL FINDINGS
    # -------------------------------------------------------------------------
    # Row 28: TABLE 3 banner (C00000 red)
    ws.merge_cells('A28:G28')
    cell = ws['A28']
    cell.value = "TABLE 3 \u2014 CRITICAL FINDINGS"
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.border = border_thin

    # Row 29: column headers
    t3_headers = ["Finding ID", "Description", "Affected Area", "Severity", "Status", "Owner", "Due Date"]
    for col, hdr in enumerate(t3_headers, start=1):
        cell = ws.cell(row=29, column=col)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Rows 30-39: 10 finding data rows (FFFFCC), pulling from Gap Analysis
    # Array formula pattern (IFERROR/INDEX/SMALL/IF) for Critical/High severity
    ga_col_map = {
        1: "A",   # Finding ID
        2: "B",   # Description
        3: "D",   # Related Requirement (Affected Area)
        4: "E",   # Severity
        5: "K",   # Status
        6: "I",   # Owner
        7: "J",   # Due Date
    }
    for finding_num in range(1, 11):
        row_num = 29 + finding_num  # rows 30-39
        a_ref = f"A{finding_num}"  # A1, A2, ... A10 for ROWS($A$1:An)
        for col_idx in range(1, 8):
            ga_col = ga_col_map[col_idx]
            formula = (
                f"=IFERROR(INDEX('Gap Analysis'!{ga_col}$6:{ga_col}$55,"
                f"SMALL(IF(('Gap Analysis'!E$6:E$55=\"Critical\")+('Gap Analysis'!E$6:E$55=\"High\"),"
                f"ROW('Gap Analysis'!A$6:A$55)-ROW('Gap Analysis'!A$6)+1),"
                f"ROWS($A$1:${a_ref})),"
                f"1),\"\")"
            )
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = formula
            cell.fill = yellow_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(name="Calibri", size=10, color="000000")

    # Rows 40-41: FFFFCC buffer rows
    for buf_row in range(40, 42):
        for col_idx in range(1, 8):
            cell = ws.cell(row=buf_row, column=col_idx)
            cell.fill = yellow_fill
            cell.border = border_thin

    # Row 42: TOTAL row
    ws.cell(row=42, column=1).value = "Total Critical/High Findings:"
    ws.cell(row=42, column=1).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=42, column=1).border = border_thin
    ws.cell(row=42, column=2).value = "=COUNTIF('Gap Analysis'!E6:E55,\"Critical\")+COUNTIF('Gap Analysis'!E6:E55,\"High\")"
    ws.cell(row=42, column=2).font = Font(name="Calibri", size=10, bold=True, color="000000")
    ws.cell(row=42, column=2).border = border_thin
    for col_idx in range(3, 8):
        ws.cell(row=42, column=col_idx).border = border_thin

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15


# ============================================================================
# SECTION 11: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create comprehensive evidence documentation sheet (gold standard)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    validations = create_base_validations(ws)

    # Row 1: A1:H1 merged, 003366, white bold, height 35
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Row 2: italic subtitle (no fill)
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "Comprehensive evidence documentation for endpoint inventory assessment (100 evidence entries)"
    cell.font = Font(name="Calibri", size=10, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3 empty

    # Row 4: 003366 headers, white bold
    headers = [
        "Evidence ID",
        "Evidence Type",
        "Description",
        "Related Requirement",
        "Related Worksheet/Device",
        "File Location",
        "Collection Date",
        "Collected By",
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Freeze at A5
    ws.freeze_panes = 'A5'

    # Row 5: F2F2F2 grey sample row, fully populated
    sample_row_er = 5
    grey_fill_er = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_er_data = [
        "EV-001",
        "Screenshot",
        "Example: Screenshot of endpoint encryption status",
        "A.8.1 Baseline Compliance",
        "Encryption Status sheet row 5",
        "/evidence/A.8.1/enc-screenshot.png",
        "20.02.2026",
        "IT Security Team",
    ]
    for col, val in enumerate(sample_er_data, start=1):
        cell = ws.cell(row=sample_row_er, column=col)
        cell.value = val
        cell.fill = grey_fill_er
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 6-105: 100 empty FFFFCC yellow rows
    start_row = 6
    for i in range(100):
        current_row = start_row + i
        fill_color = "FFFFCC"

        # Evidence ID column (no pre-fill — user enters EV-NNN)
        cell = ws.cell(row=current_row, column=1)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.border = border_thin

        # Columns 2-8: input cells
        for col in range(2, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col == 2:
                validations['evidence_type'].add(cell)

    # Summary row (COUNTA starts at row 6 to exclude sample row)
    summary_row = start_row + 101
    ws[f'A{summary_row}'].value = "Total Evidence Entries:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = '=COUNTA(C6:C105)'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18


# ============================================================================
# SECTION 11: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create gap identification and remediation tracking sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS & REMEDIATION TRACKING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:M2')
    cell = ws['A2']
    cell.value = "Gap identification, severity classification, remediation planning and tracking (50 entries)"
    apply_style(cell, styles['subheader'])

    # Freeze at A5
    ws.freeze_panes = 'A5'

    # Column headers
    headers = [
        "Gap ID",
        "Gap Description",
        "Affected Devices/Count",
        "Related Requirement",
        "Severity",
        "Risk",
        "Root Cause",
        "Remediation Plan",
        "Owner",
        "Due Date",
        "Status",
        "Budget Required",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Gap entries: row 5 = F2F2F2 sample (GAP-001), rows 6-55 = 50 FFFFCC empty
    gap_sample_data = [
        "GAP-001",
        "Example: AV/EDR agent not installed on endpoint",
        "5 endpoints (IT Department)",
        "REQ-A81-001",
        "High",
        "High — unprotected endpoints exposed to malware",
        "Imaging process does not include AV agent",
        "Update imaging template to include AV agent pre-install",
        "IT Operations",
        "28.02.2026",
        "Open",
        "CHF 0",
        "Example row — do not use for compliance calculations"
    ]
    start_row = 5
    thin_s = Side(style="thin")
    border_thin_g = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    for i in range(51):
        current_row = start_row + i
        fill_color = "F2F2F2" if i == 0 else "FFFFCC"

        # Gap ID
        cell_id = ws.cell(row=current_row, column=1)
        if i == 0:
            cell_id.value = "GAP-001"
            cell_id.font = Font(color="808080")
        else:
            cell_id.font = Font()
        cell_id.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell_id.border = border_thin_g

        # Columns 2-13
        for col in range(2, 14):
            cell = ws.cell(row=current_row, column=col)
            if i == 0:
                cell.value = gap_sample_data[col - 1]
                cell.font = Font(color="808080")
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = border_thin_g
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col == 5:
                validations['gap_severity'].add(cell)
            elif col == 11:
                validations['gap_status'].add(cell)

    # Summary
    summary_row = start_row + 53
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} GAP SUMMARY BY SEVERITY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Critical:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"Critical")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_critical'])
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "High:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"High")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_high'])
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Medium:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"Medium")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_medium'])
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Low:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E6:E55,"Low")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_low'])
    
    summary_row += 2
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} GAP SUMMARY BY STATUS"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Open:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"Open")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "In Progress:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"In Progress")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Closed:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K6:K55,"{CHECK} Closed")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 30



# ============================================================================
# SECTION 12: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create approval sign-off sheet (Gold Standard: A.8.33-34)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G9"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    # B6: Overall Compliance Rating — format as percentage
    ws["B6"].number_format = "0.0%"

    # Assessment Status DV (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections (start row 11 after gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = border
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION (GS-AS-012: plain label, no banner fill)
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 13: MAIN FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Systems Engineering Principle: "Don't fool yourself - and you are the 
    easiest person to fool." - Richard Feynman
    
    This script generates EVIDENCE-BASED endpoint security assessment,
    not cargo cult compliance checkbox theater.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.1-7-18-19.S1 - Endpoint Inventory Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Controls: A.8.1, A.8.7, A.8.18, A.8.19")
    logger.info("=" * 78)
    logger.info("\n\u1f3af Systems Engineering Approach: Evidence-Based Compliance")
    logger.info(f"{CHART} Technology-Agnostic: Works with ANY endpoint environment")
    logger.info(f"{LOCK} Audit-Ready: Comprehensive evidence collection")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("{CHECK} Workbook created with 10 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")

    logger.info("  [1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  Instructions complete")

    logger.info("  [2/10] Creating Inventory sheet...")
    create_inventory_sheet(wb["Inventory"], styles)
    logger.info("  Inventory complete (1 grey sample + 50 data rows)")

    logger.info("  [3/10] Creating Baseline Compliance sheet...")
    create_baseline_compliance_sheet(wb["Baseline Compliance"], styles)
    logger.info("  Baseline compliance assessment complete")

    logger.info("  [4/10] Creating Encryption Status sheet...")
    create_encryption_status_sheet(wb["Encryption Status"], styles)
    logger.info("  Encryption status tracking complete")

    logger.info("  [5/10] Creating Management Enrollment sheet...")
    create_management_enrollment_sheet(wb["Management Enrollment"], styles)
    logger.info("  MDM enrollment tracking complete")

    logger.info("  [6/10] Creating Capability Requirements sheet...")
    create_capability_requirements_sheet(wb["Capability Requirements"], styles)
    logger.info("  Requirements mapping complete (30 A.8.1 requirements)")

    logger.info("  [7/10] Creating Summary Dashboard sheet...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  Summary Dashboard complete (TABLE 1/2/3 + ownership/criticality KPIs)")

    logger.info("  [8/10] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  Evidence register complete (1 grey sample + 100 evidence rows)")

    logger.info("  [9/10] Creating Gap Analysis sheet...")
    create_gap_analysis_sheet(wb["Gap Analysis"], styles)
    logger.info("  Gap analysis complete (1 grey sample + 50 gap tracking rows)")

    logger.info("  [10/10] Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  Approval workflow complete")

    # Finalise data validations
    finalize_validations(wb)

    # Save workbook to WKBK/ directory
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    output_filename = f"ISMS-IMP-A.8.1-7-18-19.S1_Endpoint_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        wb.save(output_path)
        logger.info(f"SUCCESS: {output_path}")
    except Exception as e:
        logger.error(f"ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\nAssessment Sheets:")
    logger.info("  - Instructions & Legend (usage guidance, status legend)")
    logger.info("  - Inventory (1 grey sample + 50 data rows with full metadata)")
    logger.info("  - Baseline Compliance (1 grey sample + 50 per-endpoint rows)")
    logger.info("  - Encryption Status (1 grey sample + 50 encryption tracking rows)")
    logger.info("  - Management Enrollment (1 grey sample + 50 MDM enrollment rows)")
    logger.info("\nGovernance & Evidence:")
    logger.info("  - Capability Requirements (1 grey sample + 30 A.8.1 requirements)")
    logger.info("  - Summary Dashboard (TABLE 1/2/3 consolidated metrics)")
    logger.info("  - Evidence Register (1 grey sample + 100 evidence entries)")
    logger.info("  - Gap Analysis (1 grey sample + 50 gap tracking rows)")
    logger.info("  - Approval Sign-Off (3-level approval workflow)")
    logger.info("\n" + "=" * 78)
    logger.info("NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Review Instructions & Legend sheet for usage guidance")
    logger.info("  3. Export endpoint inventory from MDM/SCCM/Jamf")
    logger.info("  4. Populate Inventory sheet with endpoint data")
    logger.info("  5. Complete Baseline Compliance, Encryption Status assessments")
    logger.info("  6. Document evidence in Evidence Register (min 20 entries)")
    logger.info("  7. Identify gaps in Gap Analysis sheet")
    logger.info("  8. Obtain approvals via Approval Sign-Off workflow")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
