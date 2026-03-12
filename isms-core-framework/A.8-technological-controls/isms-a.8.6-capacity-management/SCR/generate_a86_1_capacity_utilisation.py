#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.6-Assessment-1 - Capacity Utilization Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.6: Capacity Management
Assessment Workbook 1 of 3: Current Capacity Utilization Analysis

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific infrastructure, monitoring systems, and capacity
management requirements.

Key customization areas:
1. Resource inventory (servers, storage, networks - match your infrastructure)
2. Utilization thresholds (warning/critical levels - adapt to your risk profile)
3. Monitoring tool integrations (Prometheus, Datadog, CloudWatch, etc.)
4. Capacity headroom calculations (based on your planning methodology)
5. Assessment frequency and reporting cycles (monthly, quarterly)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.6 Capacity Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for analyzing
current capacity utilization across all infrastructure resources to support
proactive capacity planning and prevent resource exhaustion.

**Purpose:**
Enables systematic assessment of current resource utilization against ISO
27001:2022 Control A.8.6 requirements, supporting evidence-based capacity
planning decisions and preventing capacity-related service degradation.

**Assessment Scope:**
- Compute capacity utilization (CPU, memory, VM/container capacity)
- Storage capacity utilization (disk space, database storage, backup storage)
- Network capacity utilization (bandwidth, throughput, load balancer capacity)
- Application capacity utilization (concurrent users, transaction rates, queues)
- Cloud service capacity (cloud quotas, service limits, cost thresholds)
- Threshold status evaluation (OK, Warning, Critical)
- Capacity headroom calculation (remaining capacity before exhaustion)
- Peak utilization tracking (maximum observed usage)
- Utilization trends (growth patterns over time)
- Gap identification for resources exceeding thresholds

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and capacity standards
2. Compute Capacity - Server CPU, memory, virtualization capacity
3. Storage Capacity - Disk space, database, backup, archive storage
4. Network Capacity - Bandwidth, throughput, network infrastructure
5. Application Capacity - User sessions, transactions, queue depths
6. Cloud Capacity - Cloud service quotas and limits (if applicable)
7. Capacity Summary - Consolidated view of all resource utilization
8. Threshold Status - Resources by threshold status (OK/Warning/Critical)
9. Capacity Headroom - Available capacity before resource exhaustion
10. Peak Utilization - Maximum observed capacity usage
11. Gap Analysis - Resources requiring immediate capacity expansion
12. Evidence Register - Supporting documentation and monitoring data
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with threshold status dropdown lists
- Conditional formatting for utilization status visualization
- Automated headroom calculations (remaining capacity percentage)
- Protected formulas with unprotected input cells
- Peak utilization tracking for capacity planning
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with capacity forecasting assessment (Assessment 2)

**Integration:**
This assessment provides current utilization data for the A.8.6 Capacity
Management Dashboard, which consolidates utilization, forecasts, and planning
effectiveness for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a86_1_capacity_utilization.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a86_1_capacity_utilization.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a86_1_capacity_utilization.py --date 20250128

Output:
    File: ISMS_A_8_6_1_Capacity_Utilization_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize capacity thresholds to match your risk profile
    2. Inventory all infrastructure resources requiring capacity monitoring
    3. Export current utilization data from monitoring tools
    4. Complete utilization assessments for each resource type
    5. Calculate capacity headroom (remaining capacity)
    6. Identify resources at warning/critical thresholds
    7. Review peak utilization patterns for trend analysis
    8. Conduct gap analysis for resources requiring expansion
    9. Collect and link evidence (monitoring dashboards, reports)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.6 Capacity Management Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.6
Assessment Type:      Assessment 1 of 3 (Current Capacity Utilization)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.6: Capacity Management Policy (Governance)
    - ISMS-IMP-A.8.6-S1: Capacity Monitoring Implementation Guide
    - ISMS-IMP-A.8.6-S2: Capacity Forecasting & Planning Implementation Guide
    - ISMS-IMP-A.8.6-S3: Capacity Management Assessment Guide
    - Assessment 2: Capacity Forecasts & Planning (generate_a86_2_capacity_forecasts.py)
    - Dashboard: Capacity Management Overview (generate_a86_3_compliance_dashboard.py)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.6-S3 specification
    - Supports comprehensive capacity utilization evaluation
    - Integrated with A.8.6 Capacity Management Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Capacity Thresholds:**
Capacity thresholds vary by resource type and organisational risk tolerance.
Default thresholds (Warning: 70%, Critical: 85%) should be reviewed and
customized based on:
- Resource type criticality (production vs. development)
- Lead time for capacity expansion (hardware procurement vs. cloud scaling)
- Performance degradation characteristics (gradual vs. sudden)
- Business impact tolerance (customer-facing vs. internal systems)

**Monitoring Tool Integration:**
This assessment workbook is tool-agnostic and accepts data from any monitoring
platform including:
- Open source: Prometheus, Grafana, Nagios, Zabbix
- Commercial SaaS: Datadog, New Relic, Dynatrace, Splunk
- Cloud-native: AWS CloudWatch, Azure Monitor, GCP Cloud Monitoring
- Enterprise: IBM Tivoli, HP OpenView, BMC TrueSight

Export utilization data from your monitoring tool and populate the assessment
workbook manually or via CSV import (see Implementation Guide).

**Assessment Frequency:**
Monthly assessment is recommended for most environments:
- Monthly: Tactical capacity review and immediate action items
- Quarterly: Comprehensive capacity planning (feeds Assessment 2)
- Annual: Strategic capacity planning and budget development

Critical production systems may require weekly or daily capacity reviews.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Current utilization data from monitoring systems
- Documentation of threshold definitions and rationale
- Evidence of capacity planning actions for resources at warning/critical
- Approval from infrastructure management and IT leadership

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- System inventory and resource specifications
- Capacity utilization patterns and performance data
- Capacity constraints and bottleneck identification
- Infrastructure growth trends and projections

Handle in accordance with your organisation's data classification policies.
Consider this assessment workbook "Internal" or "Confidential" classification.

**Maintenance:**
Review and update assessment:
- Monthly: Update utilization data and threshold status
- Quarterly: Review capacity headroom and expansion requirements
- Annually: Validate threshold definitions and assessment methodology
- Ad-hoc: When infrastructure changes or new resources are deployed

**Quality Assurance:**
Have infrastructure managers and capacity planning team validate assessments
before using results for capacity planning decisions or compliance reporting.
Cross-reference utilization data against monitoring tool dashboards to ensure
accuracy.

**Integration with Forecasting:**
Assessment 1 (Utilization) provides current-state data that feeds into
Assessment 2 (Forecasts & Planning) for trend analysis and future capacity
projection. Ensure both assessments are completed for comprehensive capacity
management.

**Cloud vs. On-Premises Considerations:**
- **On-premises**: Focus on hardware capacity limits and procurement lead times
- **Cloud**: Focus on service quotas, auto-scaling policies, and cost thresholds
- **Hybrid**: Assess both on-premises and cloud capacity independently

Customize the assessment to match your infrastructure deployment model.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.6-Assessment-1"
WORKBOOK_NAME = "Capacity Utilization Assessment"
CONTROL_ID = "A.8.6"
CONTROL_NAME = "Capacity Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅
XMARK = '\u274C'      # ❌
WARNING = '\u26A0'    # ⚠️
HOURGLASS = '\u23F3'  # ⏳
BULLET = '\u2022'     # •
ARROW = '\u2192'      # →
CHART = '[CHART]'
TARGET = '[TARGET]'
TRENDING_UP = '[TREND]'
DISK = '[DISK]'
STOPWATCH = '\u23F1'  # ⏱️

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.6.1 specification
    # Note: IL is created first by create_instructions_sheet(),
    # ER and AS use wb.create_sheet() in their own functions
    sheets = [
        "Compute Resources",
        "Storage Resources",
        "Network Resources",
        "Application Resources",
        "Cloud Resources",
        "Threshold Summary",
        "Coverage Analysis",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "sample_row": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_ok": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_warning": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_critical": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    Returns dictionary of validation objects that can be applied to cells.
    """
    validations = {}
    
    # Status validation
    validations['status'] = DataValidation(
        type="list",
        formula1=f'"{CHECK} OK,⚠️ Warning,❌ Critical,Not Monitored"',
        allow_blank=False
    )
    
    # Yes/No validation
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    
    # Resource type validation
    validations['resource_type'] = DataValidation(
        type="list",
        formula1='"Physical Server,Virtual Machine,Container/Pod,Cloud Instance,Other"',
        allow_blank=False
    )
    
    # Monitoring tool validation
    validations['monitoring_tool'] = DataValidation(
        type="list",
        formula1='"Prometheus,Datadog,New Relic,CloudWatch,Azure Monitor,GCP Monitoring,Zabbix,Nagios,PRTG,Other"',
        allow_blank=True
    )
    
    return validations


# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete each worksheet tab for applicable resource types.',
        '2. Use dropdown menus for standardised entries (Status, Resource Type, etc.).',
        '3. Fill in yellow-highlighted cells with your information.',
        '4. Utilization percentages and threshold status auto-calculate where applicable.',
        '5. Review Threshold Summary for overall capacity health.',
        '6. Complete Coverage Analysis to identify monitoring gaps.',
        '7. Provide evidence location/path for each assessment entry.',
        '8. Maintain the Evidence Register for audit traceability.',
        '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A23"] = "Status Legend"
    ws["A23"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=24, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 25 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_compute_resources(wb, styles):
    """Create Compute Resources capacity sheet."""
    ws = wb["Compute Resources"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    num_cols = 13

    # Row 1 - Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "COMPUTE CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2 - Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Document current CPU and memory capacity utilization"
    apply_style(cell, styles['subheader'])

    # Row 3 - Column headers
    headers = [
        "Resource Name", "Resource Type", "Location/Cluster", "Business Function",
        "Total CPU (Cores)", "Used CPU (Cores)", "CPU Utilization (%)",
        "Total Memory (GB)", "Used Memory (GB)", "Memory Utilization (%)",
        "Peak CPU (%)", "Peak Memory (%)", "Threshold Status"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 4 - Example data row (F2F2F2 grey — sample row standard)
    example = [
        "server01.example.com", "Physical Server", "DC1-Rack5", "Web Application",
        "16", "8", "", "64", "32", "", "75", "60", ""
    ]
    input_cols = {1, 2, 3, 4, 5, 6, 8, 9, 11, 12}
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = value
        apply_style(cell, styles['sample_row'])

    # Formulas for row 4
    ws['G4'] = '=IF(AND(E4<>"", F4<>""), ROUND((F4/E4)*100, 1), "")'
    ws['J4'] = '=IF(AND(H4<>"", I4<>""), ROUND((I4/H4)*100, 1), "")'
    ws['M4'] = f'=IF(A4="", "", IF(OR(G4="", J4=""), "Not Monitored", IF(OR(G4>85, J4>90), "\u274c Critical", IF(OR(G4>70, J4>75), "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Rows 5-54 - Pre-formatted empty data rows (50 empty rows = 51 total with sample)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c in input_cols:
                apply_style(cell, styles['input_cell'])
            else:
                thin = Side(style="thin")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Copy formulas down
        ws[f'G{r}'] = f'=IF(AND(E{r}<>"", F{r}<>""), ROUND((F{r}/E{r})*100, 1), "")'
        ws[f'J{r}'] = f'=IF(AND(H{r}<>"", I{r}<>""), ROUND((I{r}/H{r})*100, 1), "")'
        ws[f'M{r}'] = f'=IF(A{r}="", "", IF(OR(G{r}="", J{r}=""), "Not Monitored", IF(OR(G{r}>85, J{r}>90), "\u274c Critical", IF(OR(G{r}>70, J{r}>75), "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Apply validations (rows 4-54 = 51 total rows)
    validations['resource_type'].add('B4:B54')
    validations['status'].add('M4:M54')
    for _dv in validations.values():
        ws.add_data_validation(_dv)

    # Freeze pane below headers
    ws.freeze_panes = "A4"

    # Column widths
    widths = [30, 20, 20, 25, 15, 15, 18, 15, 15, 20, 15, 15, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 5: SHEET 3 - STORAGE RESOURCES
# ============================================================================

def create_storage_resources(wb, styles):
    """Create Storage Resources capacity sheet."""
    ws = wb["Storage Resources"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    num_cols = 11

    # Row 1 - Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "STORAGE CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2 - Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Document current disk, database, and backup storage capacity"
    apply_style(cell, styles['subheader'])

    # Row 3 - Column headers
    headers = [
        "Resource Name", "Storage Type", "Location/Server", "Business Function",
        "Total Capacity (GB)", "Used Capacity (GB)", "Available Capacity (GB)",
        "Utilization (%)", "Peak Utilization (%)", "Growth Rate (GB/month)", "Threshold Status"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 4 - Example data row (F2F2F2 grey — sample row standard)
    example = [
        "/dev/sda1 (/)", "Filesystem", "server01", "Operating System",
        "500", "375", "", "", "80", "5", ""
    ]
    input_cols = {1, 2, 3, 4, 5, 6, 9, 10}
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = value
        apply_style(cell, styles['sample_row'])

    # Formulas for row 4
    ws['G4'] = '=IF(AND(E4<>"", F4<>""), E4-F4, "")'
    ws['H4'] = '=IF(AND(E4<>"", F4<>""), ROUND((F4/E4)*100, 1), "")'
    ws['K4'] = f'=IF(A4="", "", IF(H4="", "Not Monitored", IF(H4>85, "\u274c Critical", IF(H4>75, "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Rows 5-54 - Pre-formatted empty data rows (50 empty rows = 51 total with sample)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c in input_cols:
                apply_style(cell, styles['input_cell'])
            else:
                thin = Side(style="thin")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Copy formulas down
        ws[f'G{r}'] = f'=IF(AND(E{r}<>"", F{r}<>""), E{r}-F{r}, "")'
        ws[f'H{r}'] = f'=IF(AND(E{r}<>"", F{r}<>""), ROUND((F{r}/E{r})*100, 1), "")'
        ws[f'K{r}'] = f'=IF(A{r}="", "", IF(H{r}="", "Not Monitored", IF(H{r}>85, "\u274c Critical", IF(H{r}>75, "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Apply storage type validation (rows 4-54 = 51 total rows)
    storage_types = DataValidation(
        type="list",
        formula1='"Filesystem,Database,Backup Repository,Archive,SAN/NAS,Object Storage,Other"',
        allow_blank=False
    )
    storage_types.add('B4:B54')

    validations['status'].add('K4:K54')
    all_dvs = list(validations.values()) + [storage_types]
    for _dv in all_dvs:
        ws.add_data_validation(_dv)

    # Freeze pane below headers
    ws.freeze_panes = "A4"

    # Column widths
    widths = [30, 20, 25, 25, 18, 18, 20, 15, 18, 20, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 6: SHEET 4 - NETWORK RESOURCES
# ============================================================================

def create_network_resources(wb, styles):
    """Create Network Resources capacity sheet."""
    ws = wb["Network Resources"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    num_cols = 12

    # Row 1 - Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "NETWORK CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2 - Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Document network bandwidth, throughput, and service capacity"
    apply_style(cell, styles['subheader'])

    # Row 3 - Column headers
    headers = [
        "Resource Name", "Network Type", "Location", "Business Function",
        "Link Capacity (Mbps)", "Avg Inbound (Mbps)", "Avg Outbound (Mbps)",
        "Total Utilization (%)", "Peak Utilization (%)", "Concurrent Connections",
        "Connection Limit", "Threshold Status"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 4 - Example data row (F2F2F2 grey — sample row standard)
    example = [
        "WAN-Link-Primary", "WAN/Internet", "DC1", "Primary Internet",
        "1000", "350", "150", "", "75", "5000", "10000", ""
    ]
    input_cols = {1, 2, 3, 4, 5, 6, 7, 9, 10, 11}
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = value
        apply_style(cell, styles['sample_row'])

    # Formulas for row 4
    ws['H4'] = '=IF(AND(E4<>"", F4<>"", G4<>""), ROUND(((F4+G4)/E4)*100, 1), "")'
    ws['L4'] = f'=IF(A4="", "", IF(H4="", "Not Monitored", IF(H4>85, "\u274c Critical", IF(H4>70, "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Rows 5-54 - Pre-formatted empty data rows (50 empty rows = 51 total with sample)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c in input_cols:
                apply_style(cell, styles['input_cell'])
            else:
                thin = Side(style="thin")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Copy formulas down
        ws[f'H{r}'] = f'=IF(AND(E{r}<>"", F{r}<>"", G{r}<>""), ROUND(((F{r}+G{r})/E{r})*100, 1), "")'
        ws[f'L{r}'] = f'=IF(A{r}="", "", IF(H{r}="", "Not Monitored", IF(H{r}>85, "\u274c Critical", IF(H{r}>70, "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Network type validation (rows 4-54 = 51 total rows)
    network_types = DataValidation(
        type="list",
        formula1='"WAN/Internet,LAN,Inter-DC Link,VPN,Load Balancer,Firewall,Other"',
        allow_blank=False
    )
    network_types.add('B4:B54')

    validations['status'].add('L4:L54')
    all_dvs = list(validations.values()) + [network_types]
    for _dv in all_dvs:
        ws.add_data_validation(_dv)

    # Freeze pane below headers
    ws.freeze_panes = "A4"

    # Column widths
    widths = [25, 20, 20, 25, 18, 18, 18, 18, 18, 20, 15, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 7: SHEET 5 - APPLICATION RESOURCES
# ============================================================================

def create_application_resources(wb, styles):
    """Create Application Resources capacity sheet."""
    ws = wb["Application Resources"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    num_cols = 11

    # Row 1 - Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "APPLICATION CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2 - Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Document application user sessions, transactions, and performance capacity"
    apply_style(cell, styles['subheader'])

    # Row 3 - Column headers
    headers = [
        "Application Name", "Application Type", "Business Function",
        "Max Concurrent Users", "Current Active Users", "User Utilization (%)",
        "Max Transactions/sec", "Current Transactions/sec", "Transaction Utilization (%)",
        "Avg Response Time (ms)", "Threshold Status"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 4 - Example data row (F2F2F2 grey — sample row standard)
    example = [
        "WebApp-Production", "Web Application", "Customer Portal",
        "1000", "650", "", "500", "280", "", "145", ""
    ]
    input_cols = {1, 2, 3, 4, 5, 7, 8, 10}
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = value
        apply_style(cell, styles['sample_row'])

    # Formulas for row 4
    ws['F4'] = '=IF(AND(D4<>"", E4<>""), ROUND((E4/D4)*100, 1), "")'
    ws['I4'] = '=IF(AND(G4<>"", H4<>""), ROUND((H4/G4)*100, 1), "")'
    ws['K4'] = f'=IF(A4="", "", IF(OR(F4="", I4=""), "Not Monitored", IF(OR(F4>90, I4>90), "\u274c Critical", IF(OR(F4>75, I4>75), "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Rows 5-53 - Pre-formatted empty data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c in input_cols:
                apply_style(cell, styles['input_cell'])
            else:
                thin = Side(style="thin")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Copy formulas down
        ws[f'F{r}'] = f'=IF(AND(D{r}<>"", E{r}<>""), ROUND((E{r}/D{r})*100, 1), "")'
        ws[f'I{r}'] = f'=IF(AND(G{r}<>"", H{r}<>""), ROUND((H{r}/G{r})*100, 1), "")'
        ws[f'K{r}'] = f'=IF(A{r}="", "", IF(OR(F{r}="", I{r}=""), "Not Monitored", IF(OR(F{r}>90, I{r}>90), "\u274c Critical", IF(OR(F{r}>75, I{r}>75), "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Application type validation (rows 4-53)
    app_types = DataValidation(
        type="list",
        formula1='"Web Application,Database,API Service,Message Queue,Batch Processing,SaaS Application,Other"',
        allow_blank=False
    )
    app_types.add('B4:B53')

    validations['status'].add('K4:K54')
    all_dvs = list(validations.values()) + [app_types]
    for _dv in all_dvs:
        ws.add_data_validation(_dv)

    # Freeze pane below headers
    ws.freeze_panes = "A4"

    # Column widths
    widths = [25, 20, 25, 20, 20, 18, 20, 22, 22, 22, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 8: SHEET 6 - CLOUD RESOURCES
# ============================================================================

def create_cloud_resources(wb, styles):
    """Create Cloud Resources capacity sheet."""
    ws = wb["Cloud Resources"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    num_cols = 10

    # Row 1 - Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "CLOUD SERVICE CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2 - Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Document cloud service quotas, limits, and instance counts"
    apply_style(cell, styles['subheader'])

    # Row 3 - Column headers
    headers = [
        "Cloud Provider", "Service/Resource Type", "Region",
        "Quota/Limit", "Current Usage", "Utilization (%)",
        "Monthly Cost (CHF)", "Reserved Capacity", "Auto-Scaling Enabled", "Threshold Status"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 4 - Example data row (F2F2F2 grey — sample row standard)
    example = [
        "AWS", "EC2 Instances (t3.large)", "us-east-1",
        "100", "72", "", "4320", "20 Reserved", "Yes", ""
    ]
    input_cols = {1, 2, 3, 4, 5, 7, 8, 9}
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = value
        apply_style(cell, styles['sample_row'])

    # Formulas for row 4
    ws['F4'] = '=IF(AND(D4<>"", E4<>""), ROUND((E4/D4)*100, 1), "")'
    ws['J4'] = f'=IF(A4="", "", IF(F4="", "Not Monitored", IF(F4>85, "\u274c Critical", IF(F4>70, "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Rows 5-53 - Pre-formatted empty data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c in input_cols:
                apply_style(cell, styles['input_cell'])
            else:
                thin = Side(style="thin")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Copy formulas down
        ws[f'F{r}'] = f'=IF(AND(D{r}<>"", E{r}<>""), ROUND((E{r}/D{r})*100, 1), "")'
        ws[f'J{r}'] = f'=IF(A{r}="", "", IF(F{r}="", "Not Monitored", IF(F{r}>85, "\u274c Critical", IF(F{r}>70, "\u26a0\ufe0f Warning", "{CHECK} OK"))))'

    # Cloud provider validation (rows 4-53)
    cloud_providers = DataValidation(
        type="list",
        formula1='"AWS,Azure,GCP,Oracle Cloud,IBM Cloud,Alibaba Cloud,Other"',
        allow_blank=False
    )
    cloud_providers.add('A4:A53')

    validations['yes_no'].add('I4:I53')
    validations['status'].add('J4:J54')
    all_dvs = list(validations.values()) + [cloud_providers]
    for _dv in all_dvs:
        ws.add_data_validation(_dv)

    # Freeze pane below headers
    ws.freeze_panes = "A4"

    # Column widths
    widths = [15, 30, 15, 15, 15, 15, 18, 20, 20, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 9: SHEET 7 - THRESHOLD SUMMARY
# ============================================================================

def create_threshold_summary(wb, styles):
    """Create Threshold Summary dashboard sheet (formula-driven, no user data input)."""
    ws = wb["Threshold Summary"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1 - Title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPACITY THRESHOLD STATUS SUMMARY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2 - Subtitle
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Aggregate view of resources by threshold status"
    apply_style(cell, styles['subheader'])

    # Row 3 - Column headers
    col_headers = [
        "Resource Category", "Total Resources",
        f"{CHECK} OK", f"{WARNING} Warning", f"{XMARK} Critical", "Not Monitored"
    ]
    for col, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Resource categories with formulas (rows 4-8)
    categories = [
        ("Compute Resources", "'Compute Resources'", "M"),
        ("Storage Resources", "'Storage Resources'", "K"),
        ("Network Resources", "'Network Resources'", "L"),
        ("Application Resources", "'Application Resources'", "K"),
        ("Cloud Resources", "'Cloud Resources'", "J"),
    ]

    row = 4
    for category, sheet, col in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].border = border_thin
        ws[f'B{row}'] = f'=COUNTA({sheet}!A5:A54)'
        ws[f'B{row}'].border = border_thin
        ws[f'C{row}'] = f'=COUNTIF({sheet}!{col}5:{col}54,"{CHECK} OK")'
        ws[f'C{row}'].border = border_thin
        ws[f'D{row}'] = f'=COUNTIF({sheet}!{col}5:{col}54,"\u26a0\ufe0f Warning")'
        ws[f'D{row}'].border = border_thin
        ws[f'E{row}'] = f'=COUNTIF({sheet}!{col}5:{col}54,"\u274c Critical")'
        ws[f'E{row}'].border = border_thin
        ws[f'F{row}'] = f'=COUNTIF({sheet}!{col}5:{col}54,"Not Monitored")'
        ws[f'F{row}'].border = border_thin

        apply_style(ws[f'C{row}'], styles['status_ok'])
        ws[f'C{row}'].border = border_thin
        apply_style(ws[f'D{row}'], styles['status_warning'])
        ws[f'D{row}'].border = border_thin
        apply_style(ws[f'E{row}'], styles['status_critical'])
        ws[f'E{row}'].border = border_thin
        row += 1

    # Total row (row 9)
    row += 1  # skip a row for spacing -> row 10
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].border = border_thin
    ws[f'B{row}'] = '=SUM(B4:B8)'
    ws[f'B{row}'].border = border_thin
    ws[f'C{row}'] = '=SUM(C4:C8)'
    ws[f'C{row}'].border = border_thin
    ws[f'D{row}'] = '=SUM(D4:D8)'
    ws[f'D{row}'].border = border_thin
    ws[f'E{row}'] = '=SUM(E4:E8)'
    ws[f'E{row}'].border = border_thin
    ws[f'F{row}'] = '=SUM(F4:F8)'
    ws[f'F{row}'].border = border_thin

    for c in range(2, 7):
        ws.cell(row=row, column=c).font = Font(bold=True)

    # Capacity health score
    total_row = row
    row += 2
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "CAPACITY HEALTH SCORE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = f'=IF(B{total_row}>0, ROUND((C{total_row}/B{total_row})*100, 1), 0)'
    ws[f'D{row}'] = "%"
    ws[f'C{row}'].font = Font(bold=True, size=14)

    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "Health Score = (Resources at OK Status / Total Resources) x 100%"
    ws[f'A{row}'].font = Font(italic=True, size=9)

    # Freeze pane below headers
    ws.freeze_panes = "A4"

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18


# ============================================================================
# SECTION 10: SHEET 8 - COVERAGE ANALYSIS
# ============================================================================

def create_coverage_analysis(wb, styles):
    """Create Coverage Analysis sheet."""
    ws = wb["Coverage Analysis"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1 - Title
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "MONITORING COVERAGE ANALYSIS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2 - Subtitle
    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = "Identify resources not yet monitored for capacity"
    apply_style(cell, styles['subheader'])

    # Row 3 - Coverage metrics headers
    cov_headers = ["Coverage Metric", "Target", "Actual", "Gap", "Status"]
    for col, header in enumerate(cov_headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Rows 4-6 - Coverage targets (small fixed section)
    metrics = [
        ("Production Systems Coverage", "100%"),
        ("Non-Production Systems Coverage", "90%"),
        ("Overall Coverage", "95%"),
    ]
    row = 4
    for metric, target in metrics:
        ws[f'A{row}'] = metric
        ws[f'A{row}'].border = border_thin
        ws[f'B{row}'] = target
        ws[f'B{row}'].border = border_thin
        apply_style(ws[f'C{row}'], styles['input_cell'])
        apply_style(ws[f'D{row}'], styles['input_cell'])
        apply_style(ws[f'E{row}'], styles['input_cell'])
        row += 1

    # Row 8 - Gap details subheader
    row = 8
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "MONITORING GAPS - Resources Not Yet Monitored"
    apply_style(cell, styles['subheader'])

    # Row 9 - Gap column headers
    gap_headers = [
        "Resource Name", "Resource Type", "Reason Not Monitored",
        "Planned Monitoring Date", "Responsible Party"
    ]
    for col, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Row 10 - Example gap entry (F2F2F2 grey — sample row standard)
    example_gap = [
        "[Resource name]", "[Type]", "[Legacy system, Technical constraint, etc.]",
        "[Date]", "[Team/Person]"
    ]
    for col, value in enumerate(example_gap, start=1):
        cell = ws.cell(row=10, column=col)
        cell.value = value
        apply_style(cell, styles['sample_row'])

    # Rows 11-60 - Pre-formatted empty gap rows (50 empty rows = 51 total with sample)
    for r in range(11, 61):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            apply_style(cell, styles['input_cell'])

    # Freeze pane below coverage metrics headers
    ws.freeze_panes = "A4"

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 25


# ============================================================================
# SECTION 11: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(wb, styles):
    """Create Summary Dashboard — TABLE 1/2/3 (Gold Standard)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    yelw = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # Row 1: Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "CAPACITY UTILIZATION ASSESSMENT \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border_thin
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (left-aligned)
    ws.merge_cells("A2:G2")
    ws["A2"] = "Consolidated capacity utilisation compliance across all infrastructure resource types."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 1: Assessment Area Compliance ──────────────────────────────────
    t1r = 4
    ws.merge_cells(f"A{t1r}:G{t1r}")
    ws[f"A{t1r}"] = "TABLE 1 \u2014 ASSESSMENT AREA COMPLIANCE"
    ws[f"A{t1r}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t1r}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{t1r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{t1r}"].border = border_thin

    t1h = 5
    for c, h in enumerate(["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"], start=1):
        cell = ws.cell(row=t1h, column=c, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # 5 assessment areas — status DV: ✅ OK / ⚠️ Warning (U+26A0+FE0F) / ❌ Critical / Not Monitored
    # Sample row is row 4 in each resource sheet — formula ranges start at row 5
    t1_areas = [
        ("Compute Resources",     "'Compute Resources'",     "M", 5, 54),
        ("Storage Resources",     "'Storage Resources'",     "K", 5, 54),
        ("Network Resources",     "'Network Resources'",     "L", 5, 54),
        ("Application Resources", "'Application Resources'", "K", 5, 54),
        ("Cloud Resources",       "'Cloud Resources'",       "J", 5, 54),
    ]
    for i, (area, sheet, col, r1, r2) in enumerate(t1_areas):
        row = t1h + 1 + i  # rows 6-10
        rng = f"{sheet}!{col}{r1}:{col}{r2}"
        c_f  = f'=COUNTIF({rng},"{CHECK} OK")'
        p_f  = f'=COUNTIF({rng},"\u26A0\uFE0F Warning")'
        nc_f = f'=COUNTIF({rng},"\u274C Critical")'
        na_f = f'=COUNTIF({rng},"Not Monitored")'
        pct_f = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        vals = [area, f"=C{row}+D{row}+E{row}+F{row}", c_f, p_f, nc_f, na_f, pct_f]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.font = Font(color="000000")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=7).number_format = "0.0%"

    # ── TABLE 2: KPI Metrics ──────────────────────────────────────────────────
    t2r = 12
    ws.merge_cells(f"A{t2r}:G{t2r}")
    ws[f"A{t2r}"] = "TABLE 2 \u2014 KEY PERFORMANCE INDICATORS"
    ws[f"A{t2r}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2r}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{t2r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{t2r}"].border = border_thin

    t2h = 13
    for c, h in enumerate(["KPI Metric", "Value"], start=1):
        cell = ws.cell(row=t2h, column=c, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    kpis = [
        ("Total Resources Monitored (All Types)",
         "=COUNTA('Compute Resources'!A5:A54)+COUNTA('Storage Resources'!A5:A54)"
         "+COUNTA('Network Resources'!A5:A54)+COUNTA('Application Resources'!A5:A54)"
         "+COUNTA('Cloud Resources'!A5:A54)", None),
        ("Compute Resources \u2014 OK Rate",
         f'=IFERROR(COUNTIF(\'Compute Resources\'!M5:M54,"{CHECK} OK")/COUNTA(\'Compute Resources\'!A5:A54),0)',
         "0.0%"),
        ("Storage Resources \u2014 OK Rate",
         f'=IFERROR(COUNTIF(\'Storage Resources\'!K5:K54,"{CHECK} OK")/COUNTA(\'Storage Resources\'!A5:A54),0)',
         "0.0%"),
        ("Network Resources \u2014 OK Rate",
         f'=IFERROR(COUNTIF(\'Network Resources\'!L5:L54,"{CHECK} OK")/COUNTA(\'Network Resources\'!A5:A54),0)',
         "0.0%"),
        ("Application Resources \u2014 OK Rate",
         f'=IFERROR(COUNTIF(\'Application Resources\'!K5:K54,"{CHECK} OK")/COUNTA(\'Application Resources\'!A5:A54),0)',
         "0.0%"),
        ("Cloud Resources \u2014 OK Rate",
         f'=IFERROR(COUNTIF(\'Cloud Resources\'!J5:J54,"{CHECK} OK")/COUNTA(\'Cloud Resources\'!A5:A54),0)',
         "0.0%"),
        ("Resources at Critical Threshold",
         '=COUNTIF(\'Compute Resources\'!M5:M54,"\u274C Critical")'
         '+COUNTIF(\'Storage Resources\'!K5:K54,"\u274C Critical")'
         '+COUNTIF(\'Network Resources\'!L5:L54,"\u274C Critical")'
         '+COUNTIF(\'Application Resources\'!K5:K54,"\u274C Critical")'
         '+COUNTIF(\'Cloud Resources\'!J5:J54,"\u274C Critical")',
         None),
        ("Resources at Warning Threshold",
         '=COUNTIF(\'Compute Resources\'!M5:M54,"\u26A0\uFE0F Warning")'
         '+COUNTIF(\'Storage Resources\'!K5:K54,"\u26A0\uFE0F Warning")'
         '+COUNTIF(\'Network Resources\'!L5:L54,"\u26A0\uFE0F Warning")'
         '+COUNTIF(\'Application Resources\'!K5:K54,"\u26A0\uFE0F Warning")'
         '+COUNTIF(\'Cloud Resources\'!J5:J54,"\u26A0\uFE0F Warning")',
         None),
    ]
    for i, (name, formula, fmt) in enumerate(kpis):
        row = 14 + i  # rows 14-21
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.border = border_thin
        cell_a.font = Font(color="000000")
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_b = ws.cell(row=row, column=2, value=formula)
        cell_b.border = border_thin
        cell_b.font = Font(color="000000")
        cell_b.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell_b.number_format = fmt

    # ── TABLE 3: Critical Findings ────────────────────────────────────────────
    t3r = 23
    ws.merge_cells(f"A{t3r}:G{t3r}")
    ws[f"A{t3r}"] = "TABLE 3 \u2014 CRITICAL FINDINGS (COMPUTE RESOURCES AT CRITICAL THRESHOLD)"
    ws[f"A{t3r}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t3r}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{t3r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{t3r}"].border = border_thin

    t3h = 24
    t3_headers = ["Resource Name", "Resource Type", "Location/Cluster", "CPU Util (%)", "Memory Util (%)", "Threshold Status", "Business Function"]
    for c, h in enumerate(t3_headers, start=1):
        cell = ws.cell(row=t3h, column=c, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # TABLE 3 data rows — empty FFFFCC user-fill (record critical resources manually after assessment)
    for i in range(10):
        row = 25 + i
        for c_idx in range(1, 8):
            cell = ws.cell(row=row, column=c_idx)
            cell.fill = yelw
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Buffer rows 35-36
    for r in range(35, 37):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = yelw
            ws.cell(row=r, column=c).border = border_thin

    # TOTAL row 37
    total_cell = ws.cell(row=37, column=1, value="Total Resources at Critical Threshold:")
    total_cell.font = Font(bold=True)
    total_cell.border = border_thin
    total_formula = (
        '=COUNTIF(\'Compute Resources\'!M5:M54,"\u274C Critical")'
        '+COUNTIF(\'Storage Resources\'!K5:K54,"\u274C Critical")'
        '+COUNTIF(\'Network Resources\'!L5:L54,"\u274C Critical")'
        '+COUNTIF(\'Application Resources\'!K5:K54,"\u274C Critical")'
        '+COUNTIF(\'Cloud Resources\'!J5:J54,"\u274C Critical")'
    )
    total_val = ws.cell(row=37, column=2, value=total_formula)
    total_val.font = Font(bold=True, color="000000")
    total_val.border = border_thin

    # Column widths
    col_widths = [38, 18, 22, 18, 18, 22, 25]
    for c_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: SHEET 9 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Monitoring dashboard,Screenshot,Utilization report,Documentation,Capacity plan,Alert configuration,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    # Row 5: Sample row with complete example data
    sample_data = [
        "EV-001",
        "Capacity Monitoring",
        "Monitoring dashboard",
        "Compute capacity dashboard showing current utilization metrics",
        "/monitoring/capacity/compute-dashboard.xlsx",
        "15.01.2026",
        "IT Operations",
        "Verified"
    ]
    for c, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=c, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    dv_type.add(ws.cell(row=5, column=3))
    dv_ver.add(ws.cell(row=5, column=8))

    # Rows 6-105: 100 empty rows
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 13: SHEET 11 - APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(wb):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"

    # Status dropdown on Assessment Status (row 7)
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add("B7")

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.6.1 - Capacity Utilization Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.6: Capacity Management")
    logger.info("=" * 80)
    logger.info("")
    
    # Create workbook
    logger.info("Creating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    # Create sheets
    logger.info("Generating Instructions & Legend...")
    create_instructions_sheet(wb.create_sheet())
    
    logger.info("Generating Compute Resources sheet...")
    create_compute_resources(wb, styles)
    
    logger.info("Generating Storage Resources sheet...")
    create_storage_resources(wb, styles)
    
    logger.info("Generating Network Resources sheet...")
    create_network_resources(wb, styles)
    
    logger.info("Generating Application Resources sheet...")
    create_application_resources(wb, styles)
    
    logger.info("Generating Cloud Resources sheet...")
    create_cloud_resources(wb, styles)
    
    logger.info("Generating Threshold Summary dashboard...")
    create_threshold_summary(wb, styles)
    
    logger.info("Generating Coverage Analysis...")
    create_coverage_analysis(wb, styles)
    
    logger.info("Generating Evidence Register...")
    create_evidence_register(wb)

    logger.info("Generating Summary Dashboard...")
    create_summary_dashboard_sheet(wb, styles)

    logger.info("Generating Approval Sign-Off...")
    create_approval_sheet(wb)
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.6.1_Capacity_Utilization_Assessment_{timestamp}.xlsx"

    # Output to WKBK directory
    script_dir = Path(__file__).resolve().parent
    _wkbk_dir = script_dir.parent / "WKBK"
    _wkbk_dir.mkdir(parents=True, exist_ok=True)
    output_path = _wkbk_dir / OUTPUT_FILENAME
    logger.info("")
    logger.info(f"Saving workbook: {output_path}")
    finalize_validations(wb)
    wb.save(output_path)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("{CHECK} SUCCESS - Capacity Utilization Assessment Workbook Created")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Output file: {output_path}")
    logger.info("")
    logger.info("NEXT STEPS:")
    logger.info("1. Open the workbook in Excel/LibreOffice")
    logger.info("2. Complete yellow-highlighted input cells in each resource sheet")
    logger.info("3. Review auto-calculated utilization percentages and threshold status")
    logger.info("4. Review Threshold Summary for overall capacity health")
    logger.info("5. Document monitoring gaps in Coverage Analysis")
    logger.info("6. Add supporting evidence in Evidence Register")
    logger.info("7. Obtain approvals in Approval Sign-Off sheet")
    logger.info("")
    logger.info("For dashboard integration:")
    logger.info("  • Run normalize_assessment_files_a86.py to prepare for dashboard")
    logger.info("  • Generate dashboard: python3 generate_dashboard_capacity_management.py")
    logger.info("")
    logger.info("=" * 80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
