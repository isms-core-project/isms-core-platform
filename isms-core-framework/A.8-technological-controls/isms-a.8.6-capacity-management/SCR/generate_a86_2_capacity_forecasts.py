#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.6-Assessment-2 - Capacity Forecasting & Planning Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.6: Capacity Management
Assessment Workbook 2 of 3: Capacity Forecasts and Planning Effectiveness

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific forecasting methodology, planning cycles, and
capacity management maturity.

Key customization areas:
1. Forecasting horizon (6, 12, 24 months - match your planning cycles)
2. Forecasting methodology (linear regression, seasonal models, growth rates)
3. Forecast accuracy thresholds (acceptable error ranges for your context)
4. Planning cycle frequency (monthly, quarterly, annual reviews)
5. Capacity expansion criteria (when to trigger procurement/provisioning)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.6 Capacity Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for capacity
forecasting, planning effectiveness evaluation, and proactive capacity
expansion planning based on trend analysis and business growth projections.

**Purpose:**
Enables systematic capacity forecasting and planning assessment against ISO
27001:2022 Control A.8.6 requirements, supporting evidence-based capacity
expansion decisions and preventing reactive "firefighting" capacity additions.

**Assessment Scope:**
- Capacity forecasting (6-month, 12-month, 24-month projections)
- Trend analysis (linear growth, seasonal patterns, anomaly detection)
- Forecasted capacity exhaustion dates (when resources will be depleted)
- Capacity expansion planning (planned additions, procurement status)
- Forecast accuracy validation (previous forecasts vs. actual utilization)
- Planning effectiveness metrics (proactive vs. reactive expansions)
- Capacity-related incidents (capacity exhaustion events, near-misses)
- Forecast confidence intervals (uncertainty ranges)
- Business growth alignment (capacity planning with business projections)
- Gap analysis for forecasting methodology improvements

**Generated Workbook Structure:**
1. Instructions & Legend - Forecasting methodology and planning guidance
2. Historical Utilization - Historical data for trend analysis (6-24 months)
3. Trend Analysis - Growth rates, seasonal patterns, regression analysis
4. 6-Month Forecast - Short-term capacity projections (tactical planning)
5. 12-Month Forecast - Medium-term capacity projections (budget planning)
6. 24-Month Forecast - Long-term capacity projections (strategic planning)
7. Capacity Exhaustion - Forecasted dates when resources will be depleted
8. Expansion Planning - Planned capacity additions with timelines and status
9. Forecast Accuracy - Validation of previous forecasts vs. actual utilization
10. Planning Effectiveness - Proactive vs. reactive capacity expansions
11. Capacity Incidents - Capacity-related outages and near-miss events
12. Business Alignment - Capacity planning aligned with business growth
13. Gap Analysis - Forecasting methodology improvements required
14. Evidence Register - Supporting documentation (forecast models, approvals)
15. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Historical utilization data import from Assessment 1 (Utilization)
- Automated trend analysis and growth rate calculations
- Multiple forecasting horizons (short, medium, long-term)
- Capacity exhaustion date calculation (time until depletion)
- Forecast accuracy metrics (MAPE, MAE, forecast bias)
- Proactive vs. reactive expansion tracking (planning effectiveness)
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with capacity management dashboard

**Integration:**
This assessment combines current utilization (Assessment 1) with forecasting
and planning data to feed the A.8.6 Capacity Management Dashboard, providing
comprehensive capacity health visibility for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a86_2_capacity_forecasts.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a86_2_capacity_forecasts.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a86_2_capacity_forecasts.py --date 20250128

Output:
    File: ISMS_A_8_6_2_Capacity_Forecasts_Planning_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Import historical utilization data from Assessment 1 (last 6-24 months)
    2. Export historical data from monitoring tools for trend analysis
    3. Perform trend analysis (linear regression, growth rate calculation)
    4. Generate capacity forecasts for 6, 12, and 24-month horizons
    5. Calculate forecasted capacity exhaustion dates
    6. Document planned capacity expansions with timelines
    7. Validate forecast accuracy by comparing previous forecasts to actuals
    8. Track proactive vs. reactive capacity expansions (planning effectiveness)
    9. Document capacity-related incidents and near-miss events
    10. Align capacity planning with business growth projections
    11. Conduct gap analysis for forecasting methodology improvements
    12. Collect and link evidence (forecast models, approval documents)
    13. Obtain stakeholder approvals
    14. Feed results into A.8.6 Capacity Management Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.6
Assessment Type:      Assessment 2 of 3 (Capacity Forecasts & Planning)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.6: Capacity Management Policy (Governance)
    - ISMS-IMP-A.8.6-S1: Capacity Monitoring Implementation Guide
    - ISMS-IMP-A.8.6-S2: Capacity Forecasting & Planning Implementation Guide
    - ISMS-IMP-A.8.6-S3: Capacity Management Assessment Guide
    - Assessment 1: Capacity Utilization (generate_a86_1_capacity_utilization.py)
    - Dashboard: Capacity Management Overview (generate_a86_3_compliance_dashboard.py)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full forecasting and planning assessment framework
    - Supports multiple forecasting horizons (6, 12, 24 months)
    - Integrated with A.8.6 Capacity Management Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Forecasting Methodology:**
Multiple forecasting approaches are supported in this framework:
- **Linear Regression**: Best for steady, predictable growth patterns
- **Growth Rate Calculation**: Useful for percentage-based growth projection
- **Seasonal Models**: For resources with recurring usage patterns
- **Business-Driven Forecasts**: Align capacity with planned business growth

Default methodology is linear regression on historical utilization data.
Customize based on your infrastructure characteristics and data patterns.

**Historical Data Requirements:**
Quality forecasts require sufficient historical data:
- **Minimum**: 6 months of historical utilization data
- **Recommended**: 12-24 months for better trend identification
- **Data Quality**: <5% missing data points, outliers removed
- **Data Granularity**: Daily or weekly aggregates for trend analysis

Export historical data from your monitoring tools before completing Assessment 2.

**Forecast Accuracy Targets:**
Forecast accuracy varies by resource type and forecasting horizon:
- **Acceptable**: Forecast within ±15% of actual utilization (MAPE ≤15%)
- **Good**: Forecast within ±10% of actual utilization (MAPE ≤10%)
- **Excellent**: Forecast within ±5% of actual utilization (MAPE ≤5%)

Short-term forecasts (6 months) should be more accurate than long-term (24 months).
Document assumptions and confidence intervals for all forecasts.

**Planning Effectiveness Metrics:**
Track proactive vs. reactive capacity expansions:
- **Proactive**: Capacity expansion planned in advance (based on forecast)
- **Reactive**: Emergency capacity addition (unplanned, due to exhaustion)
- **Target**: ≥90% proactive expansions (per ISMS-POL-A.8.6-S2)

Low proactive ratio indicates forecasting or planning process deficiencies.

**Capacity Exhaustion Calculation:**
For each resource, calculate forecasted exhaustion date:
```
Months Until Exhaustion = (100% - Current Utilization) / Monthly Growth Rate
Exhaustion Date = Current Date + Months Until Exhaustion
```

Resources with exhaustion dates within procurement lead time require immediate
capacity planning action (add to expansion planning sheet).

**Assessment Frequency:**
- **Monthly**: Review short-term forecasts (6 months), update expansion tracker
- **Quarterly**: Comprehensive forecast update (6, 12 months), planning review
- **Annually**: Long-term strategic planning (24 months), methodology review

Align assessment frequency with your organisation's planning cycles.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will expect:
- Documented forecasting methodology with justification
- Historical utilization data supporting forecasts
- Forecast accuracy validation (previous forecasts vs. actuals)
- Capacity expansion plans with approval evidence
- Proactive planning effectiveness metrics (≥90% proactive target)
- Capacity incident tracking (zero capacity-related outages target)

**Data Protection:**
Assessment workbooks contain sensitive strategic planning information:
- Capacity forecasts and infrastructure growth projections
- Capacity expansion plans and budget requirements
- Capacity-related incidents and performance issues
- Business growth projections and strategic initiatives

Handle in accordance with your organisation's data classification policies.
Consider this assessment workbook "Confidential" classification.

**Forecasting Tool Options:**
- **Excel**: Built-in FORECAST.LINEAR, TREND, GROWTH functions
- **Python**: Pandas, Scikit-learn (linear regression), Prophet (seasonal)
- **Monitoring Tools**: Prometheus, Datadog, CloudWatch built-in forecasting
- **Dedicated Tools**: Enterprise capacity planning platforms

This workbook supports manual forecasting or import from external tools.
See ISMS-IMP-A.8.6-S2 for detailed forecasting methodology guidance.

**Integration with Current Utilization:**
Assessment 2 (Forecasts) builds on Assessment 1 (Utilization):
- Import current utilization data from Assessment 1
- Use historical utilization for trend analysis
- Combine current state + forecasts for comprehensive capacity planning
- Both assessments feed the Capacity Management Dashboard

Complete Assessment 1 before Assessment 2 for optimal workflow.

**Continuous Improvement:**
Use forecast accuracy validation to improve forecasting methodology:
- Compare previous forecasts to actual observed utilization
- Calculate forecast error metrics (MAPE, MAE, bias)
- Identify systematic over-forecasting or under-forecasting
- Adjust forecasting parameters based on accuracy analysis
- Document lessons learned and methodology improvements

Quarterly forecast accuracy review recommended for continuous improvement.

**Cloud vs. On-Premises Planning:**
- **On-premises**: Plan for procurement lead times (weeks to months)
- **Cloud**: Plan for auto-scaling policies and cost thresholds
- **Hybrid**: Separate forecasts for on-premises and cloud resources

Forecasting horizons and expansion planning differ significantly between
on-premises (long lead time) and cloud (instant provisioning).

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from pathlib import Path
from datetime import datetime, timedelta
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.6-Assessment-2"
WORKBOOK_NAME = "Capacity Forecasting & Planning"
CONTROL_ID = "A.8.6"
CONTROL_NAME = "Capacity Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅
XMARK = '\u274C'      # ❌
WARNING = '\u26A0'    # ⚠️
HOURGLASS = '\u23F3'  # ⏳
BULLET = '\u2022'     # •
ARROW = '\u2192'      # →
CHART = '[CHART]'
TARGET = '[TARGET]'
TRENDING_UP = '[TREND]'
DISK = '[DISK]'
STOPWATCH = '\u23F1'  # ⏱️

# Import shared functions from Assessment 1
# Reuse: create_workbook structure, setup_styles, apply_style patterns

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Historical Utilization",
        "Trend Analysis",
        "Capacity Forecasts",
        "Capacity Exhaustion",
        "Planned Expansions",
        "Forecast Accuracy",
        "Budget Planning",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define cell styles (same as Assessment 1)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_ok": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_warning": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_critical": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style to cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def create_instructions(wb, styles):
    """Create Instructions & Legend sheet matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Instructions & Legend", 0)  # Always first sheet

    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.8.6.2  -  Capacity Forecasts and Planning Assessment\n"
        "ISO/IEC 27001:2022 - Control A.8.6: Capacity Management"
    )
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.6.2"),
        ("Assessment Area", "Capacity Forecasting and Planning"),
        ("Related Policy", "ISMS-POL-A.8.6"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border_thin
        row += 1

    row += 1
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete each worksheet tab for applicable systems/services.",
        "2. Use dropdown menus for standardised entries (Status, Remediation, etc.).",
        "3. Fill in yellow-highlighted cells with your information.",
        "4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.",
        "5. Check all applicable items in the Compliance Checklist for each section.",
        "6. Provide evidence location/path for each implementation entry.",
        "7. Summary Dashboard auto-calculates compliance statistics.",
        "8. Maintain the Evidence Register for audit traceability.",
        "9. Obtain final approval and sign-off in the Approval Sign-Off sheet.",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    legend_headers = ["Symbol", "Status", "Description"]
    legend_data = [
        ("\u2705", "Compliant", "Fully meets policy requirements"),
        ("\u26A0\uFE0F", "Partial", "Some requirements met, gaps exist"),
        ("\u274C", "Non-Compliant", "Does not meet policy requirements"),
        ("\u2014", "N/A", "Not applicable to this record"),
    ]

    row += 1
    for c, h in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, name="Calibri", size=11)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border_thin

    row += 1
    fills = {
        "Compliant": "C6EFCE",
        "Partial": "FFEB9C",
        "Non-Compliant": "FFC7CE",
    }
    for sym, status, desc in legend_data:
        ws.cell(row=row, column=1, value=sym).border = border_thin
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if status in fills:
            s.fill = PatternFill(start_color=fills[status], end_color=fills[status], fill_type="solid")
        row += 1

    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    evidence_types = [
        "Capacity utilization reports and dashboards",
        "Forecast models and trend analysis outputs",
        "Capacity expansion planning documents",
        "Budget approval records",
        "Monitoring tool screenshots / exports",
        "Vendor quotations and procurement records",
        "Historical utilization data exports",
    ]
    row += 1
    for e in evidence_types:
        ws[f"A{row}"] = f"\u2713 {e}"
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


def create_hist_utilization(wb, styles):
    """Create Historical Utilization sheet."""
    ws = wb["Historical Utilization"]
    ws.sheet_view.showGridLines = False
    num_cols = 6

    # Row 1: Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "HISTORICAL CAPACITY UTILIZATION DATA"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Import historical utilization data from monitoring tools for trend analysis (6-24 months recommended)"
    apply_style(cell, styles['subheader'])

    # Row 3: Column headers
    headers = ["Month/Date", "Resource Name", "Utilization (%)", "Peak Utilization (%)", "Notes", "Data Source"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 4: F2F2F2 sample row
    sample_data = ["01.01.2026", "server01.example.com", "65.4", "78.2", "Q1 historical data", "Prometheus"]
    for c, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 5-54: 50 empty FFFFCC data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25
    ws.freeze_panes = "A4"


def create_trend_analysis(wb, styles):
    """Create Trend Analysis sheet."""
    ws = wb["Trend Analysis"]
    ws.sheet_view.showGridLines = False
    num_cols = 7

    # Row 1: Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "CAPACITY TREND ANALYSIS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Growth rates, seasonal patterns, and regression analysis per resource"
    apply_style(cell, styles['subheader'])

    # Row 3: Column headers
    headers = ["Resource Name", "Trend Method", "Growth Rate (% per month)", "R-Squared", "Seasonal Pattern", "Notes", "Analyst"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # DataValidation for Trend Method (col B) and Seasonal Pattern (col E)
    dv_method = DataValidation(
        type="list",
        formula1='"Linear Regression,Growth Rate,Seasonal Model,Business-Driven,Manual Estimate"',
        allow_blank=True,
    )
    dv_seasonal = DataValidation(
        type="list",
        formula1='"None,Weekly,Monthly,Quarterly,Annual,Custom"',
        allow_blank=True,
    )

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 4: F2F2F2 sample row
    sample_data = ["server01.example.com", "Linear Regression", "2.3", "0.87", "None", "Steady monthly growth pattern", "IT Operations"]
    for c, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 5-54: 50 empty FFFFCC data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_method.add(ws.cell(row=r, column=2))
        dv_seasonal.add(ws.cell(row=r, column=5))

    ws.add_data_validation(dv_method)
    ws.add_data_validation(dv_seasonal)

    for width, col_letter in [(30, 'A'), (20, 'B'), (25, 'C'), (15, 'D'), (20, 'E'), (30, 'F'), (20, 'G')]:
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A4"


def create_capacity_forecasts(wb, styles):
    """Create Capacity Forecasts sheet."""
    ws = wb["Capacity Forecasts"]
    ws.sheet_view.showGridLines = False
    num_cols = 9

    # Row 1: Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "CAPACITY FORECASTS (6, 12, AND 24 MONTHS)"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Short-, medium-, and long-term capacity projections per resource (columns D-F auto-calculate)"
    apply_style(cell, styles['subheader'])

    # Row 3: Column headers
    headers = ["Resource Name", "Current Utilization (%)", "Growth Rate (%/month)", "6-Month Forecast (%)",
               "12-Month Forecast (%)", "24-Month Forecast (%)", "Confidence", "Assumptions", "Forecast Date"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # DataValidation for Confidence (col G)
    dv_confidence = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True,
    )

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 4: F2F2F2 sample row (cols A,B,C,G,H,I input; D,E,F formula-driven)
    sample_data = ["server01.example.com", "65.4", "2.3", "", "", "", "High", "Based on 12-month historical trend", "22.02.2026"]
    for c, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['D4'] = '=IF(AND(B4<>"",C4<>""),B4+(C4*6),"")'
    ws['E4'] = '=IF(AND(B4<>"",C4<>""),B4+(C4*12),"")'
    ws['F4'] = '=IF(AND(B4<>"",C4<>""),B4+(C4*24),"")'

    # Rows 5-54: 50 empty FFFFCC data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Auto-calculate forecast columns
        ws[f'D{r}'] = f'=IF(AND(B{r}<>"",C{r}<>""),B{r}+(C{r}*6),"")'
        ws[f'E{r}'] = f'=IF(AND(B{r}<>"",C{r}<>""),B{r}+(C{r}*12),"")'
        ws[f'F{r}'] = f'=IF(AND(B{r}<>"",C{r}<>""),B{r}+(C{r}*24),"")'
        dv_confidence.add(ws.cell(row=r, column=7))

    ws.add_data_validation(dv_confidence)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 15
    ws.freeze_panes = "A4"


def create_capacity_exhaustion(wb, styles):
    """Create Capacity Exhaustion sheet."""
    ws = wb["Capacity Exhaustion"]
    ws.sheet_view.showGridLines = False
    num_cols = 8

    # Row 1: Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "CAPACITY EXHAUSTION PROJECTIONS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Forecasted dates when resources will reach capacity thresholds (columns E-G auto-calculate)"
    apply_style(cell, styles['subheader'])

    # Row 3: Column headers
    headers = ["Resource Name", "Current (%)", "Threshold (%)", "Growth Rate (%/mo)",
               "Months to Threshold", "Exhaustion Date", "Urgency", "Action Required"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 4: F2F2F2 sample row (cols A,B,C,D,H input; E,F,G formula-driven)
    sample_data = ["server01.example.com", "65.4", "85", "2.3", "", "", "", "Initiate capacity expansion planning"]
    for c, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['E4'] = '=IF(AND(B4<>"",C4<>"",D4<>"",D4>0),(C4-B4)/D4,"")'
    ws['F4'] = '=IF(E4<>"",TODAY()+(E4*30),"")'
    ws['G4'] = '=IF(A4="","",IF(E4="","No Risk",IF(E4<=0,"CRITICAL",IF(E4<=3,"Immediate (0-3mo)",IF(E4<=6,"Short-term (3-6mo)",IF(E4<=12,"Medium-term (6-12mo)","Long-term (>12mo)"))))))'

    # Rows 5-54: 50 empty FFFFCC data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Formulas: Months to threshold, Exhaustion date, Urgency
        ws[f'E{r}'] = f'=IF(AND(B{r}<>"",C{r}<>"",D{r}<>"",D{r}>0),(C{r}-B{r})/D{r},"")'
        ws[f'F{r}'] = f'=IF(E{r}<>"",TODAY()+(E{r}*30),"")'
        ws[f'G{r}'] = f'=IF(A{r}="","",IF(E{r}="","No Risk",IF(E{r}<=0,"CRITICAL",IF(E{r}<=3,"Immediate (0-3mo)",IF(E{r}<=6,"Short-term (3-6mo)",IF(E{r}<=12,"Medium-term (6-12mo)","Long-term (>12mo)"))))))'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 30
    ws.freeze_panes = "A4"


def create_planned_expansions(wb, styles):
    """Create Planned Expansions sheet."""
    ws = wb["Planned Expansions"]
    ws.sheet_view.showGridLines = False
    num_cols = 10

    # Row 1: Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "PLANNED CAPACITY EXPANSIONS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Planned capacity additions with timelines, cost estimates, and approval status"
    apply_style(cell, styles['subheader'])

    # Row 3: Column headers
    headers = ["Resource Name", "Current Capacity", "Expansion Amount", "New Total Capacity",
               "Planned Date", "Lead Time (days)", "Status", "Cost Estimate", "Approval Status", "Owner"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # DataValidation for Status (col G) and Approval Status (col I)
    dv_status = DataValidation(
        type="list",
        formula1='"Planned,In Procurement,Ordered,Delivered,Installed,Completed,Cancelled"',
        allow_blank=True,
    )
    dv_approval = DataValidation(
        type="list",
        formula1='"Pending,Approved,Rejected,Deferred"',
        allow_blank=True,
    )

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 4: F2F2F2 sample row
    sample_data = ["server01.example.com", "2 TB", "2 TB", "4 TB", "30.06.2026", "90", "In Procurement", "15,000", "Approved", "IT Infrastructure"]
    for c, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 5-54: 50 empty FFFFCC data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_status.add(ws.cell(row=r, column=7))
        dv_approval.add(ws.cell(row=r, column=9))

    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_approval)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 20
    ws.freeze_panes = "A4"


def create_forecast_accuracy(wb, styles):
    """Create Forecast Accuracy sheet."""
    ws = wb["Forecast Accuracy"]
    ws.sheet_view.showGridLines = False
    num_cols = 8

    # Row 1: Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "FORECAST ACCURACY VALIDATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Validation of previous forecasts vs. actual utilization (columns E-G auto-calculate)"
    apply_style(cell, styles['subheader'])

    # Row 3: Column headers
    headers = ["Resource Name", "Forecast Date", "Forecasted Value", "Actual Value",
               "Absolute Error", "Percentage Error (%)", "Accuracy Rating", "Lessons Learned"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 4: F2F2F2 sample row (cols A,B,C,D,H input; E,F,G formula-driven)
    sample_data = ["server01.example.com", "01.09.2025", "75", "72", "", "", "", "Slight underestimate — adjust growth rate model"]
    for c, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['E4'] = '=IF(AND(C4<>"",D4<>""),ABS(C4-D4),"")'
    ws['F4'] = '=IF(AND(C4<>"",D4<>"",D4<>0),ABS((C4-D4)/D4)*100,"")'
    ws['G4'] = '=IF(F4="","",IF(F4<=5,"Excellent",IF(F4<=10,"Good",IF(F4<=15,"Acceptable","Poor"))))'

    # Rows 5-54: 50 empty FFFFCC data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Formulas: Absolute error, Percentage error, Accuracy rating
        ws[f'E{r}'] = f'=IF(AND(C{r}<>"",D{r}<>""),ABS(C{r}-D{r}),"")'
        ws[f'F{r}'] = f'=IF(AND(C{r}<>"",D{r}<>"",D{r}<>0),ABS((C{r}-D{r})/D{r})*100,"")'
        ws[f'G{r}'] = f'=IF(F{r}="","",IF(F{r}<=5,"Excellent",IF(F{r}<=10,"Good",IF(F{r}<=15,"Acceptable","Poor"))))'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 30
    ws.freeze_panes = "A4"


def create_budget_planning(wb, styles):
    """Create Budget Planning sheet."""
    ws = wb["Budget Planning"]
    ws.sheet_view.showGridLines = False
    num_cols = 7

    # Row 1: Title
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    cell = ws['A1']
    cell.value = "CAPACITY BUDGET PLANNING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    cell = ws['A2']
    cell.value = "Capital and operational expenditure planning for capacity expansions (column E auto-calculates 3-year TCO)"
    apply_style(cell, styles['subheader'])

    # Row 3: Column headers
    headers = ["Resource/Project", "Quarter", "CapEx (CHF)", "OpEx (CHF/month)", "3-Year TCO (CHF)", "Approval Status", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # DataValidation for Quarter (col B) and Approval Status (col F)
    dv_quarter = DataValidation(
        type="list",
        formula1='"Q1,Q2,Q3,Q4"',
        allow_blank=True,
    )
    dv_approval = DataValidation(
        type="list",
        formula1='"Pending,Approved,Rejected,Deferred"',
        allow_blank=True,
    )

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 4: F2F2F2 sample row (cols A,B,C,D,F,G input; E formula-driven)
    sample_data = ["Storage Expansion - server01", "Q2", "15000", "500", "", "Approved", "NAS upgrade for production storage"]
    for c, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['E4'] = '=IF(AND(C4<>"",D4<>""),C4+(D4*12*3),"")'

    # Rows 5-54: 50 empty FFFFCC data rows
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = input_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # 3-year TCO formula
        ws[f'E{r}'] = f'=IF(AND(C{r}<>"",D{r}<>""),C{r}+(D{r}*12*3),"")'
        dv_quarter.add(ws.cell(row=r, column=2))
        dv_approval.add(ws.cell(row=r, column=6))

    ws.add_data_validation(dv_quarter)
    ws.add_data_validation(dv_approval)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30
    ws.freeze_panes = "A4"



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete each worksheet tab for applicable systems/services.',
        '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).',
        '3. Fill in yellow-highlighted cells with your information.',
        '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.',
        '5. Check all applicable items in the Compliance Checklist for each section.',
        '6. Provide evidence location/path for each implementation entry.',
        '7. Summary Dashboard auto-calculates compliance statistics.',
        '8. Maintain the Evidence Register for audit traceability.',
        '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A23"] = "Status Legend"
    ws["A23"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=24, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 25 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(wb, styles):
    """Create Summary Dashboard — TABLE 1/2/3 (Gold Standard)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    yelw = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # Row 1: Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "CAPACITY FORECASTS AND PLANNING \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border_thin
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (left-aligned)
    ws.merge_cells("A2:G2")
    ws["A2"] = "Consolidated capacity forecasting and planning effectiveness across all assessment areas."
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── TABLE 1: Assessment Area Compliance ──────────────────────────────────
    t1r = 4
    ws.merge_cells(f"A{t1r}:G{t1r}")
    ws[f"A{t1r}"] = "TABLE 1 \u2014 ASSESSMENT AREA COMPLIANCE"
    ws[f"A{t1r}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t1r}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{t1r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{t1r}"].border = border_thin

    t1h = 5
    for c, h in enumerate(["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"], start=1):
        cell = ws.cell(row=t1h, column=c, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # TABLE 1 data rows (6-10) — sample row is row 4 in each data sheet, formula ranges start at row 5
    # Row 6: Capacity Forecasts — Confidence DV (High/Medium/Low) in col G
    row = 6
    cf_rng = "'Capacity Forecasts'!G5:G54"
    vals_6 = [
        "Capacity Forecasts (Confidence)",
        f"=C{row}+D{row}+E{row}+F{row}",
        f'=COUNTIF({cf_rng},"High")',
        f'=COUNTIF({cf_rng},"Medium")',
        f'=COUNTIF({cf_rng},"Low")',
        "=0",
        f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))",
    ]
    for c_idx, val in enumerate(vals_6, start=1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = Font(color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=7).number_format = "0.0%"

    # Row 7: Expansion Approval — col I of Planned Expansions (Approved/Pending/Rejected/Deferred)
    row = 7
    ea_rng = "'Planned Expansions'!I5:I54"
    vals_7 = [
        "Expansion Approval Status",
        f"=C{row}+D{row}+E{row}+F{row}",
        f'=COUNTIF({ea_rng},"Approved")',
        f'=COUNTIF({ea_rng},"Pending")',
        f'=COUNTIF({ea_rng},"Rejected")',
        f'=COUNTIF({ea_rng},"Deferred")',
        f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))",
    ]
    for c_idx, val in enumerate(vals_7, start=1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = Font(color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=7).number_format = "0.0%"

    # Row 8: Expansion Implementation — col G of Planned Expansions (Completed/In Progress/Planned/Cancelled)
    row = 8
    ei_rng = "'Planned Expansions'!G5:G54"
    partial_8 = (
        f'=COUNTIF({ei_rng},"In Procurement")+COUNTIF({ei_rng},"Ordered")'
        f'+COUNTIF({ei_rng},"Delivered")+COUNTIF({ei_rng},"Installed")'
    )
    vals_8 = [
        "Expansion Implementation Status",
        f"=C{row}+D{row}+E{row}+F{row}",
        f'=COUNTIF({ei_rng},"Completed")',
        partial_8,
        f'=COUNTIF({ei_rng},"Planned")',
        f'=COUNTIF({ei_rng},"Cancelled")',
        f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))",
    ]
    for c_idx, val in enumerate(vals_8, start=1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = Font(color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=7).number_format = "0.0%"

    # Row 9: Forecast Accuracy — col G of Forecast Accuracy (Excellent/Good/Acceptable/Poor)
    row = 9
    fa_rng = "'Forecast Accuracy'!G5:G54"
    compliant_9 = f'=COUNTIF({fa_rng},"Excellent")+COUNTIF({fa_rng},"Good")'
    vals_9 = [
        "Forecast Accuracy",
        f"=C{row}+D{row}+E{row}+F{row}",
        compliant_9,
        f'=COUNTIF({fa_rng},"Acceptable")',
        f'=COUNTIF({fa_rng},"Poor")',
        "=0",
        f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))",
    ]
    for c_idx, val in enumerate(vals_9, start=1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = Font(color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=7).number_format = "0.0%"

    # Row 10: Budget Approval — col F of Budget Planning (Approved/Pending/Rejected/Deferred)
    row = 10
    bp_rng = "'Budget Planning'!F5:F54"
    vals_10 = [
        "Budget Approval Status",
        f"=C{row}+D{row}+E{row}+F{row}",
        f'=COUNTIF({bp_rng},"Approved")',
        f'=COUNTIF({bp_rng},"Pending")',
        f'=COUNTIF({bp_rng},"Rejected")',
        f'=COUNTIF({bp_rng},"Deferred")',
        f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))",
    ]
    for c_idx, val in enumerate(vals_10, start=1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = Font(color="000000")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=7).number_format = "0.0%"

    # ── TABLE 2: KPI Metrics ──────────────────────────────────────────────────
    t2r = 12
    ws.merge_cells(f"A{t2r}:G{t2r}")
    ws[f"A{t2r}"] = "TABLE 2 \u2014 KEY PERFORMANCE INDICATORS"
    ws[f"A{t2r}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2r}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{t2r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{t2r}"].border = border_thin

    t2h = 13
    for c, h in enumerate(["KPI Metric", "Value"], start=1):
        cell = ws.cell(row=t2h, column=c, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    kpis = [
        ("Historical Records Entered",
         "=COUNTA('Historical Utilization'!A5:A54)",
         None),
        ("Resources with Growth Trends Documented",
         "=COUNTA('Trend Analysis'!A5:A54)",
         None),
        ("High Confidence Forecasts Rate",
         "=IFERROR(COUNTIF('Capacity Forecasts'!G5:G54,\"High\")/COUNTA('Capacity Forecasts'!G5:G54),0)",
         "0.0%"),
        ("Resources with Critical Exhaustion Risk",
         "=COUNTIF('Capacity Exhaustion'!G5:G54,\"CRITICAL\")",
         None),
        ("Resources with Immediate Exhaustion Risk (0\u20133 months)",
         "=COUNTIF('Capacity Exhaustion'!G5:G54,\"Immediate (0-3mo)\")",
         None),
        ("Planned Expansions Approved",
         "=COUNTIF('Planned Expansions'!I5:I54,\"Approved\")",
         None),
        ("Forecast Accuracy \u2014 Excellent/Good",
         "=COUNTIF('Forecast Accuracy'!G5:G54,\"Excellent\")+COUNTIF('Forecast Accuracy'!G5:G54,\"Good\")",
         None),
        ("Budget Items Approved",
         "=COUNTIF('Budget Planning'!F5:F54,\"Approved\")",
         None),
    ]
    for i, (name, formula, fmt) in enumerate(kpis):
        row = 14 + i  # rows 14-21
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.border = border_thin
        cell_a.font = Font(color="000000")
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_b = ws.cell(row=row, column=2, value=formula)
        cell_b.border = border_thin
        cell_b.font = Font(color="000000")
        cell_b.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell_b.number_format = fmt

    # ── TABLE 3: Critical Findings ────────────────────────────────────────────
    t3r = 23
    ws.merge_cells(f"A{t3r}:G{t3r}")
    ws[f"A{t3r}"] = "TABLE 3 \u2014 CRITICAL EXHAUSTION RISKS (CRITICAL OR IMMEDIATE 0\u20133 MONTHS)"
    ws[f"A{t3r}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t3r}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{t3r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{t3r}"].border = border_thin

    t3h = 24
    t3_headers = ["Resource Name", "Current (%)", "Threshold (%)", "Months to Threshold", "Exhaustion Date", "Urgency", "Action Required"]
    for c, h in enumerate(t3_headers, start=1):
        cell = ws.cell(row=t3h, column=c, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # TABLE 3 data rows — empty FFFFCC user-fill (record critical exhaustion risks manually)
    for i in range(10):
        row = 25 + i
        for c_idx in range(1, 8):
            cell = ws.cell(row=row, column=c_idx)
            cell.fill = yelw
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Buffer rows 35-36
    for r in range(35, 37):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = yelw
            ws.cell(row=r, column=c).border = border_thin

    # TOTAL row 37
    total_cell = ws.cell(row=37, column=1, value="Total Critical/Immediate Exhaustion Risks:")
    total_cell.font = Font(bold=True)
    total_cell.border = border_thin
    total_val = ws.cell(
        row=37, column=2,
        value='=COUNTIF(\'Capacity Exhaustion\'!G5:G54,"CRITICAL")+COUNTIF(\'Capacity Exhaustion\'!G5:G54,"Immediate (0-3mo)")'
    )
    total_val.font = Font(bold=True, color="000000")
    total_val.border = border_thin

    # Column widths
    col_widths = [38, 16, 16, 22, 18, 22, 35]
    for c_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = "A3"


def create_evidence_register(wb):
    """Create Evidence Register matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Capacity report,Screenshot,Monitoring export,Documentation,Vendor spec,Budget approval,Forecast model,Compliance report,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    validations = [dv_type, dv_ver]

    # Row 5: Sample row with complete example data
    sample_data = [
        "EV-001",
        "Capacity Forecasting",
        "Capacity report",
        "12-month capacity forecast model with growth projections",
        "/capacity/forecasts/q1-2026-forecast.xlsx",
        "15.01.2026",
        "Capacity Planning Team",
        "Verified"
    ]
    for c, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=c, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    dv_type.add(ws.cell(row=5, column=3))
    dv_ver.add(ws.cell(row=5, column=8))

    # Rows 6-105: 100 empty rows (101 total with header, 100 data rows)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    for _dv in validations:
        ws.add_data_validation(_dv)

    ws.freeze_panes = "A5"


def create_approval_sheet(wb):
    """Create Approval Sign-Off matching gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"

    # Status dropdown on Assessment Status (row 7)
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    dv_status.add("B7")

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_dec)

    ws.freeze_panes = "A3"
def main():
    """Main execution function."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.6.2 - Capacity Forecasts and Planning Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.6: Capacity Management")
    logger.info("=" * 80)
    logger.info("")
    
    logger.info("Creating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    logger.info("Generating Instructions...")
    create_instructions_sheet(wb.create_sheet("Instructions & Legend", 0))
    
    logger.info("Generating Historical Utilization...")
    create_hist_utilization(wb, styles)
    
    logger.info("Generating Trend Analysis...")
    create_trend_analysis(wb, styles)
    
    logger.info("Generating Capacity Forecasts...")
    create_capacity_forecasts(wb, styles)
    
    logger.info("Generating Capacity Exhaustion...")
    create_capacity_exhaustion(wb, styles)
    
    logger.info("Generating Planned Expansions...")
    create_planned_expansions(wb, styles)
    
    logger.info("Generating Forecast Accuracy...")
    create_forecast_accuracy(wb, styles)
    
    logger.info("Generating Budget Planning...")
    create_budget_planning(wb, styles)
    
    logger.info("Generating Evidence Register...")
    create_evidence_register(wb)

    logger.info("Generating Summary Dashboard...")
    create_summary_dashboard_sheet(wb, styles)

    logger.info("Generating Approval Sign-Off...")
    create_approval_sheet(wb)
    
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.6.2_Capacity_Forecasts_Planning_{timestamp}.xlsx"

    output_path = _wkbk_dir / OUTPUT_FILENAME
    logger.info("")
    logger.info(f"Saving workbook: {output_path}")
    finalize_validations(wb)
    wb.save(output_path)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("SUCCESS - Capacity Forecasts and Planning Assessment Workbook Created")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Output file: {output_path}")
    logger.info("")
    logger.info("NEXT STEPS:")
    logger.info("1. Open workbook in Excel/LibreOffice")
    logger.info("2. Import historical utilization data (Historical_Utilization sheet)")
    logger.info("3. Perform trend analysis and calculate growth rates")
    logger.info("4. Develop 6, 12, and 24-month capacity forecasts")
    logger.info("5. Calculate capacity exhaustion dates")
    logger.info("6. Plan capacity expansions with budget impact")
    logger.info("7. Validate forecast accuracy (if previous forecasts available)")
    logger.info("8. Obtain approvals (IT Director, CFO)")
    logger.info("")
    logger.info("For dashboard integration:")
    logger.info("  • Run normalize_assessment_files_a86.py after completing assessments")
    logger.info("  • Generate dashboard: python3 generate_dashboard_capacity_management.py")
    logger.info("")
    logger.info("=" * 80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
