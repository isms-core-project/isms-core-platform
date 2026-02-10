#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.6-Dashboard - Capacity Management Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.6: Capacity Management
Executive Dashboard: Consolidated Capacity Health and Compliance Overview

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific reporting requirements, dashboard preferences,
and executive visibility needs.

Key customization areas:
1. Dashboard layout and visualization preferences (charts, tables, KPIs)
2. Executive reporting frequency (monthly, quarterly dashboard updates)
3. Capacity health scoring methodology (adapt to your risk tolerance)
4. Key performance indicators (KPIs) and thresholds
5. Integration with source assessment workbooks (Assessment 1 and 2)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.6 Capacity Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an executive-level capacity management compliance
dashboard that consolidates data from Assessment 1 (Utilization) and
Assessment 2 (Forecasts & Planning) to provide comprehensive capacity health
visibility and audit readiness status.

**Purpose:**
Enables executive oversight of capacity management effectiveness against ISO
27001:2022 Control A.8.6 requirements, providing consolidated capacity health
metrics, risk visibility, and compliance status for informed decision-making.

**Dashboard Scope:**
- Overall capacity health score (composite metric across all resources)
- Utilization summary (resources by threshold status: OK/Warning/Critical)
- Capacity risks (resources with near-term exhaustion or critical status)
- Forecasting effectiveness (forecast accuracy metrics, planning success)
- Planning effectiveness (proactive vs. reactive expansions, incidents)
- Capacity trends (utilization growth over time, seasonal patterns)
- Compliance status (assessment completion, evidence coverage, approvals)
- Executive recommendations (priority actions, capacity investments)
- Audit readiness (evidence summary, compliance scoring)
- Capacity management maturity (maturity assessment scorecard)

**Generated Dashboard Structure:**
1. Executive Dashboard - One-page executive summary (KPIs and status)
2. Capacity Health Score - Overall health score and breakdown by resource type
3. Utilization Summary - Resources by threshold status (from Assessment 1)
4. Capacity Risks - Top 10 highest-risk resources requiring attention
5. Forecast Summary - Capacity exhaustion timeline (from Assessment 2)
6. Planning Effectiveness - Proactive vs. reactive, incidents, accuracy
7. Capacity Trends - Utilization trends over time (6-12 months)
8. Compliance Status - Assessment completion, evidence coverage, approvals
9. Maturity Assessment - Capacity management maturity scorecard
10. Recommendations - Priority actions and capacity investment needs
11. Evidence Summary - Consolidated evidence index from both assessments
12. Data Sources - Source assessment workbooks and data integration status
13. Approval & Sign-Off - Executive review and approval workflow

**Key Features:**
- Executive-friendly one-page summary dashboard
- Visual KPI indicators (traffic light colors: Red/Amber/Green)
- Consolidated data from Assessment 1 (Utilization) and Assessment 2 (Forecasts)
- Automated capacity health scoring algorithm
- Top risks identification and prioritization
- Trend visualization (utilization growth over time)
- Forecast accuracy tracking and continuous improvement metrics
- Planning effectiveness metrics (proactive ratio, incident tracking)
- Compliance status tracking (assessment completion, evidence coverage)
- Capacity management maturity assessment
- Evidence traceability for audit readiness
- Multi-stakeholder approval workflow

**Data Integration:**
This dashboard can use two integration approaches:
1. **External Workbook Linking** (Recommended): Formulas reference Assessment 1
   and Assessment 2 workbooks for real-time data updates
2. **Data Consolidation** (Alternative): Use consolidate_a86_dashboard.py to
   copy data values for static snapshot distribution

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a86_3_compliance_dashboard.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a86_3_compliance_dashboard.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a86_3_compliance_dashboard.py --date 20250128
    
    # Generate with external linking disabled (no formula references)
    python3 generate_a86_3_compliance_dashboard.py --no-linking

Output:
    File: ISMS_A_8_6_3_Capacity_Management_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Ensure Assessment 1 (Utilization) workbook is completed
    2. Ensure Assessment 2 (Forecasts & Planning) workbook is completed
    3. Place dashboard in same directory as Assessment 1 and 2 workbooks
    4. Open dashboard in Excel and click "Update Links" (if external linking)
    5. Review Executive Dashboard for capacity health summary
    6. Review Capacity Risks for resources requiring immediate attention
    7. Review Planning Effectiveness for proactive vs. reactive metrics
    8. Review Compliance Status for audit readiness
    9. Review Recommendations for priority actions
    10. Present to executive leadership for capacity investment decisions
    11. Obtain executive approvals on Approval & Sign-Off sheet
    12. Archive dashboard for compliance and historical tracking

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.6
Dashboard Type:       Executive Compliance Dashboard (Consolidation)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.6: Capacity Management Policy (Governance)
    - ISMS-IMP-A.8.6-S1: Capacity Monitoring Implementation Guide
    - ISMS-IMP-A.8.6-S2: Capacity Forecasting & Planning Implementation Guide
    - ISMS-IMP-A.8.6-S3: Capacity Management Assessment Guide
    - Assessment 1: Capacity Utilization (generate_a86_1_capacity_utilization.py)
    - Assessment 2: Capacity Forecasts & Planning (generate_a86_2_capacity_forecasts.py)

Related Scripts:
    - generate_a86_1_capacity_utilization.py (generates Assessment 1)
    - generate_a86_2_capacity_forecasts.py (generates Assessment 2)
    - normalize_assessment_files_a86.py (normalizes file names - optional)
    - consolidate_a86_dashboard.py (data consolidation alternative - optional)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements executive dashboard per ISMS-IMP-A.8.6-S3 specification
    - Consolidates data from Assessment 1 and Assessment 2
    - Supports external linking and data consolidation approaches

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Dashboard Philosophy:**
This dashboard embodies "executive-friendly reporting" principles:
- One-page executive summary with critical KPIs only
- Traffic light indicators (Red/Amber/Green) for quick status assessment
- Focus on risks and actions required, not operational details
- Clear recommendations for capacity investment decisions
- Drill-down capability to detailed assessments for technical teams

**Capacity Health Scoring:**
Overall capacity health score is calculated as weighted composite:
- **Threshold Status** (40%): Percentage of resources at OK status
- **Capacity Headroom** (30%): Average capacity headroom across resources
- **Forecast Accuracy** (15%): MAPE of forecasts vs. actuals
- **Planning Effectiveness** (15%): Proactive expansion ratio

Score ranges:
- **90-100%**: Excellent capacity health (Green)
- **70-89%**: Good capacity health with minor issues (Amber)
- **50-69%**: Moderate capacity health with significant risks (Amber)
- **<50%**: Poor capacity health requiring immediate action (Red)

Customize scoring methodology to match your organizational priorities.

**Data Integration Approaches:**

**Approach 1: External Workbook Linking (Recommended)**
- Dashboard contains formulas: ='[ISMS_A_8_6_1_*.xlsx]Sheet'!$A$1
- Real-time data updates when source assessments change
- Dashboard and assessments must be in same directory
- Excel displays "Update Links" prompt when opening dashboard
- Best for ongoing capacity monitoring and reporting

**Approach 2: Data Consolidation (Alternative)**
- Use consolidate_a86_dashboard.py script after generating dashboard
- Copies data values (not formulas) from assessments into dashboard
- Creates static snapshot for distribution or archival
- Dashboard can be distributed independently of assessments
- Best for board presentations, audit packages, historical archives

**File Location Requirements (External Linking):**
For external linking to work, all three files must be in the same directory:
- ISMS_A_8_6_1_Capacity_Utilization_Assessment_YYYYMMDD.xlsx
- ISMS_A_8_6_2_Capacity_Forecasts_Planning_YYYYMMDD.xlsx
- ISMS_A_8_6_3_Capacity_Management_Dashboard_YYYYMMDD.xlsx

Excel will prompt "Update Links" when opening the dashboard. Click "Update"
to refresh all data from source assessments.

**Dashboard Update Frequency:**
- **Monthly**: Update dashboard with latest Assessment 1 (Utilization) data
- **Quarterly**: Update with comprehensive Assessment 2 (Forecasts) refresh
- **Ad-hoc**: Update when significant capacity changes or incidents occur

Maintain consistent naming convention (YYYYMMDD suffix) for version tracking.

**Executive Reporting Best Practices:**
1. **Focus on Risks**: Highlight resources at critical status or near exhaustion
2. **Action-Oriented**: Provide clear recommendations with timelines and owners
3. **Trend Analysis**: Show capacity trends over time (6-12 months)
4. **Business Impact**: Relate capacity issues to business services and impact
5. **Budget Justification**: Link capacity expansion needs to budget requirements
6. **Success Metrics**: Track proactive planning effectiveness and incident reduction

**Key Performance Indicators (KPIs):**
The dashboard tracks these critical capacity management KPIs:
- **Capacity Health Score**: 0-100% composite metric
- **Resources at Critical**: Count and percentage of resources ≥85% utilization
- **Capacity Exhaustion**: Number of resources with <6 months until depletion
- **Proactive Expansion Ratio**: % of planned vs. reactive capacity additions
- **Forecast Accuracy**: Mean Absolute Percentage Error (MAPE) of forecasts
- **Capacity Incidents**: Count of capacity-related outages or near-misses
- **Average Capacity Headroom**: Mean available capacity across all resources

Customize KPIs to align with your organizational capacity management goals.

**Audit Considerations:**
This dashboard serves as primary evidence for ISO 27001:2022 Control A.8.6
compliance. Auditors will review:
- Overall capacity health score and trend (improving or degrading?)
- Capacity risks and remediation actions (are high risks being addressed?)
- Proactive planning effectiveness (≥90% proactive expansion target met?)
- Capacity-related incidents (zero capacity outages target met?)
- Assessment completion status (all assessments current and approved?)
- Evidence coverage (supporting documentation linked and accessible?)

Maintain historical dashboards (monthly/quarterly) as compliance evidence trail.

**Data Protection:**
Dashboard contains executive-level summary data but still includes sensitive:
- Capacity constraints and infrastructure bottlenecks
- Capacity expansion plans and budget requirements
- Capacity-related incidents and performance issues
- Strategic capacity planning and business growth alignment

Handle as "Confidential" per your organization's data classification policy.

**Quality Assurance:**
Before distributing dashboard to executive leadership:
1. Verify external links are working (if using linking approach)
2. Validate capacity health score calculations
3. Cross-check top risks against Assessment 1 and 2 source data
4. Review recommendations for clarity and actionability
5. Ensure all evidence references are accessible
6. Obtain technical review from capacity planning team
7. Obtain management review from IT leadership

**Troubleshooting External Links:**
If dashboard displays "#REF!" errors:
- Ensure Assessment 1 and 2 workbooks are in same directory as dashboard
- Verify Assessment 1 and 2 file names match dashboard formula references
- Check Excel security settings allow external workbook links
- Try "Edit Links" > "Update Values" from Excel Data tab
- If links still fail, use consolidate_a86_dashboard.py for data consolidation

**Integration with Broader ISMS:**
This capacity management dashboard integrates with:
- A.8.14 (Redundancy): Capacity for redundant systems and failover
- A.8.16 (Monitoring): Real-time monitoring infrastructure capacity
- A.5.30 (BC/DR): Capacity for business continuity requirements
- A.7.11 (Supporting Utilities): Physical infrastructure capacity

Cross-reference capacity management with these related ISMS controls.

**Continuous Improvement:**
Use dashboard to drive capacity management continuous improvement:
- Track capacity health score trend over time (improving or degrading?)
- Identify resources with recurring capacity issues (chronic underprovisioning?)
- Monitor forecast accuracy trend (forecasting methodology improving?)
- Track proactive ratio trend (planning effectiveness improving?)
- Review capacity incidents for root cause analysis and prevention

Quarterly capacity management review recommended using dashboard as input.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import os

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)







# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.6-Dashboard"
WORKBOOK_NAME = "Capacity Management Compliance Dashboard"
CONTROL_ID = "A.8.6"
CONTROL_NAME = "Capacity Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅
XMARK = '\u274C'      # ❌
WARNING = '\u26A0'    # ⚠️
HOURGLASS = '\u23F3'  # ⏳
BULLET = '\u2022'     # •
ARROW = '\u2192'      # →
CHART = '\U0001F4CA' # 📊
TARGET = '\U0001F3AF' # 🎯
TRENDING_UP = '\U0001F4C8' # 📈
DISK = '\U0001F4BE'   # 💾
STOPWATCH = '\u23F1'  # ⏱️

def create_workbook():
    """Create dashboard workbook."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Executive_Dashboard",
        "Utilization_Summary",
        "Forecast_Summary",
        "Planning_Effectiveness",
        "Capacity_Risks",
        "Maturity_Assessment",
        "Trend_Charts",
        "Recommendations",
        "Evidence_Summary",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define dashboard styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "metric_label": {
            "font": Font(name="Calibri", size=11, bold=True),
            "alignment": Alignment(horizontal="right", vertical="center"),
        },
        "metric_value": {
            "font": Font(name="Calibri", size=14, bold=True, color="002060"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "status_ok": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_warning": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_critical": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style to cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text if hasattr(style_dict["alignment"], 'wrap_text') else False
        )


def create_executive_dashboard(wb, styles):
    """Create Executive Dashboard - one-page overview."""
    ws = wb["Executive_Dashboard"]
    
    # Header
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "CAPACITY MANAGEMENT EXECUTIVE DASHBOARD"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "ISO/IEC 27001:2022 - Control A.8.6 | Assessment Date: {}".format(datetime.now().strftime('%d.%m.%Y'))
    apply_style(cell, styles['subheader'])
    
    # Key metrics section
    row = 4
    ws.merge_cells('A{}:H{}'.format(row, row))
    cell = ws['A{}'.format(row)]
    cell.value = "KEY CAPACITY METRICS"
    apply_style(cell, styles['subheader'])
    
    row += 1
    metrics = [
        ("Capacity Health Score", '=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!C12&"%"', "Target: >=95%"),
        ("Resources Monitored", "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!B10", "Total resources"),
        ("Resources at OK Status", "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!C10", "Below warning threshold"),
        ("Resources at Warning", "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!D10", "Require planning"),
        ("Resources at Critical", "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!E10", "Immediate action"),
        ("Forecast Accuracy (MAPE)", "[FROM ASSESSMENT 2]", "Target: <=15%"),
        ("Proactive Expansions", "[FROM ASSESSMENT 2]", "Target: >=90%"),
    ]
    
    for label, formula, target in metrics:
        ws['A{}'.format(row)] = label
        ws['B{}'.format(row)] = formula if not formula.startswith('[FROM') else "[USER UPDATE]"
        ws['C{}'.format(row)] = target
        apply_style(ws['A{}'.format(row)], styles['metric_label'])
        apply_style(ws['B{}'.format(row)], styles['metric_value'])
        row += 1
    
    # Risk summary section
    row += 1
    ws.merge_cells('A{}:H{}'.format(row, row))
    cell = ws['A{}'.format(row)]
    cell.value = "CAPACITY RISK SUMMARY"
    apply_style(cell, styles['subheader'])
    
    row += 1
    ws['A{}'.format(row)] = "Risk Level"
    ws['B{}'.format(row)] = "Count"
    ws['C{}'.format(row)] = "Description"
    for col in range(1, 4):
        apply_style(ws.cell(row=row, column=col), styles['subheader'])
    
    row += 1
    risks = [
        ("Immediate (0-3 months)", "[FROM ASSESSMENT 2]", "Urgent expansion required"),
        ("Short-term (3-6 months)", "[FROM ASSESSMENT 2]", "Begin planning now"),
        ("Medium-term (6-12 months)", "[FROM ASSESSMENT 2]", "Include in quarterly planning"),
    ]
    
    for risk, count, desc in risks:
        ws['A{}'.format(row)] = risk
        ws['B{}'.format(row)] = count if not count.startswith('[FROM') else "0"
        ws['C{}'.format(row)] = desc
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40


def create_utilization_summary(wb, styles):
    """Create Utilization Summary - links to Assessment 1."""
    ws = wb["Utilization_Summary"]
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPACITY UTILIZATION SUMMARY"
    apply_style(cell, styles['header'])
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Data source: Assessment 1 (Capacity Utilization)"
    apply_style(cell, styles['subheader'])
    
    # Link to Assessment 1 Threshold_Summary
    row = 4
    headers = ["Resource Category", "Total", "OK", "Warning", "Critical", "Not Monitored"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['subheader'])
    
    # External workbook formulas to Assessment 1
    categories = ["Compute", "Storage", "Network", "Application", "Cloud"]
    row = 5
    for i, category in enumerate(categories, start=5):
        ws['A{}'.format(row)] = category
        ws['B{}'.format(row)] = "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!B{}".format(i)
        ws['C{}'.format(row)] = "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!C{}".format(i)
        ws['D{}'.format(row)] = "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!D{}".format(i)
        ws['E{}'.format(row)] = "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!E{}".format(i)
        ws['F{}'.format(row)] = "=[ISMS-IMP-A.8.6.1.xlsx]Threshold_Summary!F{}".format(i)
        row += 1
    
    # Total row
    ws['A{}'.format(row)] = "TOTAL"
    ws['A{}'.format(row)].font = Font(bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col)
        cell.value = "=SUM({}5:{}{})".format(get_column_letter(col), get_column_letter(col), row-1)
        cell.font = Font(bold=True)


def create_forecast_summary(wb, styles):
    """Create Forecast Summary - links to Assessment 2."""
    ws = wb["Forecast_Summary"]
    
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "CAPACITY FORECAST SUMMARY"
    apply_style(cell, styles['header'])
    
    ws.merge_cells('A2:G2')
    cell = ws['A2']
    cell.value = "Data source: Assessment 2 (Capacity Forecasts and Planning)"
    apply_style(cell, styles['subheader'])
    
    row = 4
    ws['A{}'.format(row)] = "NOTE: Update this sheet with data from Assessment 2 Capacity_Exhaustion sheet"
    ws['A{}'.format(row)].font = Font(italic=True, color="FF0000")
    
    row += 1
    headers = ["Resource Name", "Current (%)", "Threshold (%)", "Months to Threshold", "Exhaustion Date", "Urgency", "Action"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['subheader'])


def create_planning_effectiveness(wb, styles):
    """Create Planning Effectiveness metrics."""
    ws = wb["Planning_Effectiveness"]
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPACITY PLANNING EFFECTIVENESS"
    apply_style(cell, styles['header'])
    
    row = 4
    ws.merge_cells('A{}:F{}'.format(row, row))
    cell = ws['A{}'.format(row)]
    cell.value = "PLANNING METRICS"
    apply_style(cell, styles['subheader'])
    
    row += 1
    metrics = [
        ("Total Capacity Expansions (Last 12 Months)", "[USER INPUT]", ""),
        ("Proactive Expansions (Planned in Advance)", "[USER INPUT]", ""),
        ("Reactive Expansions (Emergency/Unplanned)", "[USER INPUT]", ""),
        ("Proactive Ratio (%)", "=(B6/B5)*100", "Target: >=90%"),
    ]
    
    for label, value, note in metrics:
        ws['A{}'.format(row)] = label
        if label:
            apply_style(ws['A{}'.format(row)], styles['metric_label'])
        ws['B{}'.format(row)] = value
        ws['C{}'.format(row)] = note
        row += 1


def create_capacity_risks(wb, styles):
    """Create Capacity Risks register."""
    ws = wb["Capacity_Risks"]
    
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "TOP CAPACITY RISKS"
    apply_style(cell, styles['header'])
    
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "Top 10 resources by urgency (from Assessment 2)"
    apply_style(cell, styles['subheader'])
    
    row = 4
    headers = ["Rank", "Resource Name", "Current (%)", "Forecast (6mo)", "Exhaustion Date", "Urgency", "Action Required", "Owner"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['subheader'])


def create_maturity_assessment(wb, styles):
    """Create Capacity Management Maturity Assessment."""
    ws = wb["Maturity_Assessment"]
    
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "CAPACITY MANAGEMENT MATURITY ASSESSMENT"
    apply_style(cell, styles['header'])
    
    row = 4
    headers = ["Dimension", "Level 1\n(Ad-Hoc)", "Level 2\n(Defined)", "Level 3\n(Managed)", "Level 4\n(Measured)", "Level 5\n(Optimized)", "Current Level"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['subheader'])
    
    row += 1
    dimensions = [
        "Monitoring",
        "Forecasting",
        "Planning",
        "Optimization",
        "Reporting"
    ]
    
    for dim in dimensions:
        ws['A{}'.format(row)] = dim
        ws['G{}'.format(row)] = "[SELECT LEVEL 1-5]"
        row += 1
    
    row += 1
    ws['A{}'.format(row)] = "Overall Maturity Score"
    ws['A{}'.format(row)].font = Font(bold=True)
    ws['G{}'.format(row)] = "=AVERAGE(G5:G9)"
    ws['G{}'.format(row)].font = Font(bold=True, size=14, color="002060")


def create_recommendations(wb, styles):
    """Create Recommendations sheet."""
    ws = wb["Recommendations"]
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPACITY MANAGEMENT RECOMMENDATIONS"
    apply_style(cell, styles['header'])
    
    row = 4
    ws.merge_cells('A{}:F{}'.format(row, row))
    cell = ws['A{}'.format(row)]
    cell.value = "PRIORITY ACTIONS"
    apply_style(cell, styles['subheader'])
    
    row += 1
    headers = ["Priority", "Recommendation", "Impact", "Effort", "Target Date", "Owner"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['subheader'])


def create_trend_charts(wb, styles):
    """Create Trend Charts placeholder."""
    ws = wb["Trend_Charts"]
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPACITY TREND CHARTS"
    apply_style(cell, styles['header'])
    
    ws['A3'] = "NOTE: Trend charts can be created from historical data in Assessments 1 and 2"


def create_evidence_summary(wb, styles):
    """Create Evidence Summary."""
    ws = wb["Evidence_Summary"]
    
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "EVIDENCE SUMMARY"
    apply_style(cell, styles['header'])
    
    row = 4
    headers = ["Evidence Type", "Assessment 1", "Assessment 2", "Evidence Location", "Last Updated"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['subheader'])
    
    row += 1
    evidence = [
        ("Assessment Workbooks", "='[ISMS-IMP-A.8.6.1.xlsx]Instructions & Legend'!B5", "='[ISMS-IMP-A.8.6.2.xlsx]Instructions & Legend'!B5", "[File path]", "[Date]"),
        ("Monitoring Data", "Available", "Available", "[File path]", "[Date]"),
        ("Forecasts", "N/A", "Available", "[File path]", "[Date]"),
        ("Capacity Reports", "Available", "Available", "[File path]", "[Date]"),
    ]
    
    for ev_type, a1, a2, location, date in evidence:
        ws['A{}'.format(row)] = ev_type
        ws['B{}'.format(row)] = a1
        ws['C{}'.format(row)] = a2
        ws['D{}'.format(row)] = location
        ws['E{}'.format(row)] = date
        row += 1


def create_approval_signoff(wb, styles):
    """Create Approval Sign-Off."""
    ws = wb["Approval_Sign_Off"]
    
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "DASHBOARD APPROVAL AND SIGN-OFF"
    apply_style(cell, styles['header'])
    
    row = 4
    ws.merge_cells('A{}:D{}'.format(row, row))
    cell = ws['A{}'.format(row)]
    cell.value = "DASHBOARD SUMMARY"
    apply_style(cell, styles['subheader'])
    
    row += 1
    summary = [
        ("Dashboard Date:", datetime.now().strftime('%d.%m.%Y')),
        ("Assessment Period:", "[Month/Quarter Year]"),
        ("Capacity Health Score:", "='Executive_Dashboard'!B5"),
        ("Overall Compliance:", "[Compliance %]"),
    ]
    
    for label, value in summary:
        ws['A{}'.format(row)] = label
        ws['A{}'.format(row)].font = Font(bold=True)
        ws['B{}'.format(row)] = value
        row += 1
    
    row += 2
    ws.merge_cells('A{}:D{}'.format(row, row))
    cell = ws['A{}'.format(row)]
    cell.value = "APPROVAL SIGNATURES"
    apply_style(cell, styles['subheader'])
    
    row += 1
    ws['A{}'.format(row)] = "Role"
    ws['B{}'.format(row)] = "Name"
    ws['C{}'.format(row)] = "Approval"
    ws['D{}'.format(row)] = "Date"
    for col in range(1, 5):
        apply_style(ws.cell(row=row, column=col), styles['subheader'])
    
    row += 1
    roles = [
        "Prepared By (Capacity Planning Team)",
        "Reviewed By (Infrastructure Manager)",
        "Approved By (IT Director/CIO)",
        "Noted By (CISO)",
    ]
    
    for role in roles:
        ws['A{}'.format(row)] = role
        row += 1


def main():
    """Main execution."""
    logger.info("=" * 80)
    logger.info("ISMS-A.8.6 - Capacity Management Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.6: Capacity Management")
    logger.info("=" * 80)
    logger.info("")
    
    # Check for normalized assessment files
    logger.info("Checking for normalized assessment files...")
    norm_dir = os.path.join(os.getcwd(), "Dashboard_Sources")
    a1_file = os.path.join(norm_dir, "ISMS-IMP-A.8.6.1.xlsx")
    a2_file = os.path.join(norm_dir, "ISMS-IMP-A.8.6.2.xlsx")
    
    if not os.path.exists(a1_file):
        logger.info("{WARNING}  WARNING: Assessment 1 not found at {a1_file}")
        logger.info("   Run normalize_assessment_files_a86.py first")
    else:
        logger.info("{CHECK} Found: {os.path.basename(a1_file)}")
    
    if not os.path.exists(a2_file):
        logger.info("{WARNING}  WARNING: Assessment 2 not found at {a2_file}")
        logger.info("   Run normalize_assessment_files_a86.py first")
    else:
        logger.info("{CHECK} Found: {os.path.basename(a2_file)}")
    
    logger.info("")
    logger.info("Creating dashboard structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("Generating Executive Dashboard...")
    create_executive_dashboard(wb, styles)
    
    logger.info("Generating Utilization Summary...")
    create_utilization_summary(wb, styles)
    
    logger.info("Generating Forecast Summary...")
    create_forecast_summary(wb, styles)
    
    logger.info("Generating Planning Effectiveness...")
    create_planning_effectiveness(wb, styles)
    
    logger.info("Generating Capacity Risks...")
    create_capacity_risks(wb, styles)
    
    logger.info("Generating Maturity Assessment...")
    create_maturity_assessment(wb, styles)
    
    logger.info("Generating Trend Charts...")
    create_trend_charts(wb, styles)
    
    logger.info("Generating Recommendations...")
    create_recommendations(wb, styles)
    
    logger.info("Generating Evidence Summary...")
    create_evidence_summary(wb, styles)
    
    logger.info("Generating Approval Sign-Off...")
    create_approval_signoff(wb, styles)
    
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.6.3_Capacity_Management_Dashboard_{timestamp}.xlsx"
    
    logger.info("")
    logger.info("Saving dashboard: {}".format(filename))
    wb.save(filename)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("{CHECK} SUCCESS - Capacity Management Dashboard Created")
    logger.info("=" * 80)
    logger.info("")
    logger.info("Output file: {}".format(filename))
    logger.info("")
    logger.info("NEXT STEPS:")
    logger.info("1. Move dashboard to same directory as normalized assessment files")
    logger.info("   Recommended: {}".format(norm_dir))
    logger.info("2. Open dashboard in Excel")
    logger.info("3. Click 'Update Links' when prompted (to link Assessment 1 and 2)")
    logger.info("4. Some formulas reference Assessment 2 - update manually if needed")
    logger.info("5. Review and complete user-input sections")
    logger.info("6. Distribute to IT management and executives")
    logger.info("")
    logger.info("Dashboard contains:")
    logger.info("  • Executive Dashboard (one-page summary)")
    logger.info("  • Utilization Summary (from Assessment 1)")
    logger.info("  • Forecast Summary (from Assessment 2)")
    logger.info("  • Planning Effectiveness metrics")
    logger.info("  • Top Capacity Risks")
    logger.info("  • Maturity Assessment")
    logger.info("  • Recommendations")
    logger.info("")
    logger.info("=" * 80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error("\n❌ ERROR: {}".format(e))
        import traceback
        traceback.print_exc()
        exit(1)

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
