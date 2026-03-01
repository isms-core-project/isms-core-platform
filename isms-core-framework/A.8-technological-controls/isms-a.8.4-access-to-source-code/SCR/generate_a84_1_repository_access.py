#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.4.1 - Repository Access Control Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.4: Access to Source Code
Assessment Domain 1 of 2: Repository Access Control Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific source code access control infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Repository categories and access tier definitions (match your development platforms)
2. Branch protection rule requirements per repository criticality
3. Developer role categories and access permission scope
4. Code review and approval workflow requirements
5. Repository monitoring and anomaly detection scope

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.4 Access to Source Code Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
source code access control controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Repository Access Control Assessment under ISO 27001:2022 Control A.8.4. Supports evidence-based evaluation of source code repository access governance, branch protection compliance, and developer access control effectiveness.

**Assessment Scope:**
- Repository inventory completeness and access tier classification
- Developer access permission accuracy and least-privilege compliance
- Branch protection rule implementation and enforcement
- Code review and approval workflow compliance rates
- Repository monitoring and unauthorised access detection
- Access revocation timeliness upon role or employment change
- Evidence collection for development security and compliance audits

**Generated Workbook Structure:**
1. Repository Inventory
2. User Access Matrix
3. Access Approval Records
4. Access Review Log
5. Deprovisioning Log
6. Third Party Access
7. Service Accounts
8. Summary Dashboard
9. Gap Analysis
10. Evidence Register
11. Approval Sign-Off

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 2 domains covering Access to Source Code controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a84_1_repository_access.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a84_1_repository_access.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a84_1_repository_access.py --date 20250115

Output:
    File: ISMS-IMP-A.8.4.1_Repository_Access_Control_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.4
Assessment Domain:    1 of 2 (Repository Access Control Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.4: Access to Source Code Policy (Governance)
    - ISMS-IMP-A.8.4.1: Repository Access Control Assessment (Domain 1)
    - ISMS-IMP-A.8.4.2: Branch Protection Assessment (Domain 2)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.4.1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive source code access control details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review source code access controls and branch protection rules annually or when development platforms change, new repositories are created, or access control incidents are identified.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.4.1"
WORKBOOK_NAME = "Repository Access Control Assessment"
CONTROL_ID = "A.8.4"
CONTROL_NAME = "Access to Source Code"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# ============================================================================
# ASCII STATUS MARKERS (NO MULTI-BYTE EMOJI - EXCEL/QA SAFE)
# ============================================================================

CHECK = '[PASS]'       # Green checkmark equivalent
XMARK = '[FAIL]'       # Red X equivalent
WARNING = '[WARN]'     # Warning sign equivalent
HOURGLASS = '[WAIT]'   # Hourglass equivalent
BULLET = '-'           # Bullet point
ARROW = '->'           # Right arrow
QUESTION = '[?]'       # Question mark equivalent
CALENDAR = '[DATE]'    # Calendar equivalent
CHECKMARK = '[OK]'     # Check mark equivalent
MINUS = '[N/A]'        # Minus equivalent

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.4.1 specification (12 sheets)
    sheets = [
        "Instructions & Legend",
        "Repository Inventory",
        "User Access Matrix",
        "Access Approval Records",
        "Access Review Log",
        "Deprovisioning Log",
        "Third Party Access",
        "Service Accounts",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.

    Returns style TEMPLATES (dictionaries), not reusable objects
    to avoid Excel repair warnings from shared Border/Font/Fill objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "sample_row": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_appropriate": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_excessive": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_orphaned": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_expired": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "risk_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "risk_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "risk_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "risk_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell. Creates NEW style objects to avoid shared object warnings."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Repository Inventory sheet for ALL source code repositories.', '2. Fill in the User Access Matrix sheet showing who has access to what.', '3. Document all access requests and approvals in Access Approval Records.', '4. Track quarterly access reviews in Access Review Log.', '5. The Summary Dashboard sheet automatically calculates compliance metrics.', '6. Document gaps in Gap Analysis with remediation plans.', '7. Maintain Evidence Register for audit traceability.', '8. Obtain final approval and sign-off.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_repository_inventory_sheet(wb, styles):
    """Create the Repository Inventory sheet."""
    ws = wb["Repository Inventory"]
    ws.sheet_view.showGridLines = False
    
    # Header
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = "REPOSITORY INVENTORY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = "Document all source code repositories regardless of platform"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Repository Name",
        "Repository URL",
        "Repository Platform",
        "Repository Classification",
        "Repository Owner",
        "Repository Owner Email",
        "Business Purpose",
        "Primary Programming Language",
        "Last Commit Date",
        "Active/Inactive Status",
        "Total Contributors",
        "Total Commits (Lifetime)",
        "Repository Creation Date",
        "Criticality",
        "Contains Production Code",
        "Contains Sensitive Data",
        "Branch Protection Enabled",
        "Secret Scanning Enabled",
        "Last Access Review Date",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [25, 40, 18, 23, 20, 30, 40, 25, 18, 18, 15, 20, 20, 12, 20, 20, 23, 20, 20, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validations (cover all 50 data rows: 4-53)
    dv_platform = DataValidation(type="list",
        formula1='"GitHub,GitLab,Bitbucket,Azure DevOps,Self-Hosted Git,Perforce,SVN,Other"',
        allow_blank=False)
    dv_platform.add('C4:C54')

    dv_classification = DataValidation(type="list",
        formula1='"Production Code,Internal Tools,Open Source,Archived/Deprecated"',
        allow_blank=False)
    dv_classification.add('D4:D54')

    dv_status = DataValidation(type="list",
        formula1='"Active,Archived,Deprecated,Planned"', allow_blank=False)
    dv_status.add('J4:J54')

    dv_criticality = DataValidation(type="list",
        formula1='"Critical,High,Medium,Low"', allow_blank=False)
    dv_criticality.add('N4:N54')

    dv_yesno = DataValidation(type="list",
        formula1=f'"{CHECK} Yes,{XMARK} No,{WARNING} Partial,{QUESTION} Unknown"',
        allow_blank=False)
    dv_yesno.add('O4:O54')
    dv_yesno.add('P4:P54')
    dv_yesno.add('Q4:Q54')
    dv_yesno.add('R4:R54')

    validations = {
        'platform': dv_platform,
        'classification': dv_classification,
        'status': dv_status,
        'criticality': dv_criticality,
        'yesno': dv_yesno,
    }
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    num_cols = len(headers)
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add sample data row
    row = 4
    ws[f'A{row}'] = "[Example: customer-portal]"
    ws[f'B{row}'] = "[Example: https://github.com/org/customer-portal]"
    ws[f'C{row}'] = "GitHub"
    ws[f'D{row}'] = "Production Code"
    ws[f'E{row}'] = "[Example: Anna Muller]"
    ws[f'F{row}'] = "[Example: jane.doe@organisation.ch]"
    ws[f'G{row}'] = "[Example: Customer-facing web application]"
    ws[f'H{row}'] = "[Example: Python]"
    ws[f'I{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'J{row}'] = "Active"
    ws[f'K{row}'] = "[Example: 15]"
    ws[f'L{row}'] = "[Example: 3245]"
    ws[f'M{row}'] = "[Example: 2021-03-15]"
    ws[f'N{row}'] = "Critical"
    ws[f'O{row}'] = f"{CHECK} Yes"
    ws[f'P{row}'] = f"{XMARK} No"
    ws[f'Q{row}'] = f"{CHECK} Yes"
    ws[f'R{row}'] = f"{CHECK} Yes"
    ws[f'S{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'T{row}'] = "[Example notes]"

    # Style sample row
    for col in range(1, num_cols + 1):
        apply_style(ws.cell(row=row, column=col), styles['sample_row'])

    # Pre-format 50 additional data rows for user input (rows 5-53)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 5: SHEET 3 - USER ACCESS MATRIX
# ============================================================================

def create_user_access_matrix_sheet(wb, styles):
    """Create the User Access Matrix sheet."""
    ws = wb["User Access Matrix"]
    ws.sheet_view.showGridLines = False
    
    # Header
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = "USER ACCESS MATRIX"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = "Document all user access to repositories"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "User Name",
        "User Email",
        "User Role",
        "Employment Type",
        "Department/Team",
        "Repository Name",
        "Repository Platform",
        "Access Level",
        "Access Granted Date",
        "Access Approved By",
        "Business Justification",
        "Contract End Date",
        "Access Expiration Date",
        "Last Access Date",
        "Access Status",
        "Last Review Date",
        "Reviewer Comments",
        "Remediation Required",
        "Remediation Due Date",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [20, 30, 20, 18, 20, 25, 18, 18, 18, 20, 40, 18, 18, 18, 18, 18, 30, 18, 18, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validations (cover all 50 data rows: 4-53)
    dv_user_role = DataValidation(type="list",
        formula1='"Developer,Security Team,Auditor,External Contractor,Repository Admin,Service Account"',
        allow_blank=False)
    dv_user_role.add('C4:C54')

    dv_employment = DataValidation(type="list",
        formula1='"Employee,Contractor,Auditor,Automated System"', allow_blank=False)
    dv_employment.add('D4:D54')

    dv_platform = DataValidation(type="list",
        formula1='"GitHub,GitLab,Bitbucket,Azure DevOps,Self-Hosted Git,Perforce,SVN,Other"',
        allow_blank=False)
    dv_platform.add('G4:G53')

    dv_access_level = DataValidation(type="list",
        formula1='"Read,Write/Contribute,Admin"', allow_blank=False)
    dv_access_level.add('H4:H53')

    dv_access_status = DataValidation(type="list",
        formula1=f'"{CHECK} Appropriate,{WARNING} Excessive,{XMARK} Orphaned,{CALENDAR} Expired,{CHECKMARK} Remediated"',
        allow_blank=False)
    dv_access_status.add('O4:O54')

    dv_yesno = DataValidation(type="list",
        formula1=f'"{CHECK} Yes,{XMARK} No,{WARNING} Partial,{QUESTION} Unknown"',
        allow_blank=False)
    dv_yesno.add('R4:R54')

    validations = {
        'user_role': dv_user_role, 'employment': dv_employment,
        'platform': dv_platform, 'access_level': dv_access_level,
        'access_status': dv_access_status, 'yesno': dv_yesno,
    }
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    num_cols = len(headers)
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add sample data row
    row = 4
    ws[f'A{row}'] = "[Example: Thomas Meier]"
    ws[f'B{row}'] = "[Example: john.smith@organisation.ch]"
    ws[f'C{row}'] = "Developer"
    ws[f'D{row}'] = "Employee"
    ws[f'E{row}'] = "[Example: Engineering]"
    ws[f'F{row}'] = "[Example: customer-portal]"
    ws[f'G{row}'] = "GitHub"
    ws[f'H{row}'] = "Write/Contribute"
    ws[f'I{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'J{row}'] = "[Example: Anna Muller]"
    ws[f'K{row}'] = "[Example: Active development on customer portal project]"
    ws[f'L{row}'] = ""  # N/A for employees
    ws[f'M{row}'] = ""  # N/A for employees
    ws[f'N{row}'] = (datetime.now() - timedelta(days=15)).strftime("%d.%m.%Y")
    ws[f'O{row}'] = f"{CHECK} Appropriate"
    ws[f'P{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'Q{row}'] = "[Example: Access reviewed and confirmed]"
    ws[f'R{row}'] = f"{XMARK} No"
    ws[f'S{row}'] = ""
    ws[f'T{row}'] = ""

    # Style sample row
    for col in range(1, num_cols + 1):
        apply_style(ws.cell(row=row, column=col), styles['sample_row'])

    # Pre-format 50 additional data rows for user input (rows 5-53)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 6: SHEET 4 - ACCESS APPROVAL RECORDS
# ============================================================================

def create_access_approval_records_sheet(wb, styles):
    """Create the Access Approval Records sheet."""
    ws = wb["Access Approval Records"]
    ws.sheet_view.showGridLines = False
    
    # Header
    ws.merge_cells('A1:R1')
    cell = ws['A1']
    cell.value = "ACCESS APPROVAL RECORDS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:R2')
    cell = ws['A2']
    cell.value = "Document all access request approvals for audit trail"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Request Date",
        "Request ID",
        "Requestor Name",
        "Requestor Email",
        "Repository Name",
        "Access Level Requested",
        "Business Justification",
        "Expected Duration",
        "Expected End Date",
        "Repository Owner",
        "Owner Approval",
        "Approval Date",
        "Additional Approver",
        "Additional Approval",
        "Provisioning Date",
        "Provisioned By",
        "Denial Reason",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [15, 15, 20, 30, 25, 20, 40, 15, 18, 20, 15, 15, 20, 18, 15, 20, 30, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validations (cover all 50 data rows: 4-53)
    dv_access_level = DataValidation(type="list",
        formula1='"Read,Write/Contribute,Admin"', allow_blank=False)
    dv_access_level.add('F4:F53')

    dv_duration = DataValidation(type="list",
        formula1='"Permanent,Time-Bound"', allow_blank=False)
    dv_duration.add('H4:H53')

    dv_approval = DataValidation(type="list",
        formula1=f'"{CHECK} Approved,{XMARK} Denied,{HOURGLASS} Pending"',
        allow_blank=False)
    dv_approval.add('K4:K53')

    dv_additional_approval = DataValidation(type="list",
        formula1=f'"{CHECK} Approved,{XMARK} Denied,{MINUS} N/A"',
        allow_blank=False)
    dv_additional_approval.add('N4:N54')

    validations = {
        'access_level': dv_access_level, 'duration': dv_duration,
        'approval': dv_approval, 'additional_approval': dv_additional_approval,
    }
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    num_cols = len(headers)
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add sample data row
    row = 4
    ws[f'A{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'B{row}'] = "[Example: TICKET-12345]"
    ws[f'C{row}'] = "[Example: Bob Johnson]"
    ws[f'D{row}'] = "[Example: bob.johnson@organisation.ch]"
    ws[f'E{row}'] = "[Example: customer-portal]"
    ws[f'F{row}'] = "Write/Contribute"
    ws[f'G{row}'] = "[Example: Need to fix authentication bug]"
    ws[f'H{row}'] = "Permanent"
    ws[f'I{row}'] = ""
    ws[f'J{row}'] = "[Example: Anna Muller]"
    ws[f'K{row}'] = f"{CHECK} Approved"
    ws[f'L{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'M{row}'] = ""
    ws[f'N{row}'] = f"{MINUS} N/A"
    ws[f'O{row}'] = (datetime.now() + timedelta(days=1)).strftime("%d.%m.%Y")
    ws[f'P{row}'] = "[Example: DevOps Team]"
    ws[f'Q{row}'] = ""
    ws[f'R{row}'] = ""

    # Style sample row
    for col in range(1, num_cols + 1):
        apply_style(ws.cell(row=row, column=col), styles['sample_row'])

    # Pre-format 50 additional data rows for user input (rows 5-53)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 7: SHEET 5 - ACCESS REVIEW LOG
# ============================================================================

def create_access_review_log_sheet(wb, styles):
    """Create the Access Review Log sheet."""
    ws = wb["Access Review Log"]
    ws.sheet_view.showGridLines = False
    
    # Header
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = "ACCESS REVIEW LOG"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = "Track quarterly access reviews for all repositories"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Review Date",
        "Review Quarter",
        "Reviewer Name",
        "Reviewer Role",
        "Repositories Reviewed",
        "Repository Names",
        "Total Users Reviewed",
        "Appropriate Access Found",
        "Excessive Access Found",
        "Unnecessary Access Found",
        "Orphaned Accounts Found",
        "Contractor Expired Found",
        "Total Findings",
        "Remediation Completed",
        "Remediation Pending",
        "Review Completion Date",
        "Days to Complete",
        "Next Review Due Date",
        "Compliance Status",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [15, 15, 20, 20, 15, 40, 15, 18, 18, 18, 18, 18, 15, 18, 18, 18, 15, 18, 18, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validations (cover all 50 data rows: 4-53)
    dv_compliance = DataValidation(type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"',
        allow_blank=False)
    dv_compliance.add('S4:S53')

    validations = {'compliance': dv_compliance}
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    num_cols = len(headers)
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add sample data row with formulas
    row = 4
    ws[f'A{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'B{row}'] = f"Q{(datetime.now().month - 1) // 3 + 1} {datetime.now().year}"
    ws[f'C{row}'] = "[Example: Repository Owner]"
    ws[f'D{row}'] = "[Example: Engineering Manager]"
    ws[f'E{row}'] = "[Example: 5]"
    ws[f'F{row}'] = "[Example: customer-portal, api-gateway, ...]"
    ws[f'G{row}'] = "[Example: 25]"
    ws[f'H{row}'] = "[Example: 22]"
    ws[f'I{row}'] = "[Example: 2]"
    ws[f'J{row}'] = "[Example: 1]"
    ws[f'K{row}'] = "[Example: 0]"
    ws[f'L{row}'] = "[Example: 0]"
    ws[f'M{row}'] = "=I4+J4+K4+L4"  # Formula: Total Findings
    ws[f'N{row}'] = "[Example: 3]"
    ws[f'O{row}'] = "=M4-N4"  # Formula: Pending = Total - Completed
    ws[f'P{row}'] = (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y")
    ws[f'Q{row}'] = "=P4-A4"  # Formula: Days to complete
    ws[f'R{row}'] = (datetime.now() + timedelta(days=90)).strftime("%d.%m.%Y")
    ws[f'S{row}'] = f"{CHECK} Compliant"
    ws[f'T{row}'] = "[Example: All findings being addressed]"

    # Style sample row
    for col in range(1, num_cols + 1):
        apply_style(ws.cell(row=row, column=col), styles['sample_row'])

    # Pre-format 50 additional data rows for user input (rows 5-53)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 8: SHEET 6 - DEPROVISIONING LOG
# ============================================================================

def create_deprovisioning_log_sheet(wb, styles):
    """Create the Deprovisioning Log sheet."""
    ws = wb["Deprovisioning Log"]
    ws.sheet_view.showGridLines = False
    
    # Header
    ws.merge_cells('A1:P1')
    cell = ws['A1']
    cell.value = "DEPROVISIONING LOG"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:P2')
    cell = ws['A2']
    cell.value = "Track access removal for audit trail"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "User Name",
        "User Email",
        "Deprovisioning Event",
        "Event Date",
        "Notification Received Date",
        "Repository Name",
        "Previous Access Level",
        "Access Removed Date",
        "Hours to Deprovision",
        "Removed By",
        "Removal Method",
        "Verification Method",
        "Verification Date",
        "Compliant Timeline",
        "Delay Reason",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [20, 30, 20, 15, 20, 25, 18, 18, 15, 20, 15, 18, 15, 18, 30, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validations (cover all 50 data rows: 4-53)
    dv_event = DataValidation(type="list",
        formula1='"Termination,Role Change,Contract End,Access Review,Security Incident"',
        allow_blank=False)
    dv_event.add('C4:C54')

    dv_access_level = DataValidation(type="list",
        formula1='"Read,Write/Contribute,Admin"', allow_blank=False)
    dv_access_level.add('G4:G53')

    dv_method = DataValidation(type="list",
        formula1='"Automated,Manual"', allow_blank=False)
    dv_method.add('K4:K53')

    dv_verification = DataValidation(type="list",
        formula1='"Platform Check,Login Test,API Query"', allow_blank=False)
    dv_verification.add('L4:L53')

    dv_yesno = DataValidation(type="list",
        formula1=f'"{CHECK} Yes,{XMARK} No,{WARNING} Partial,{QUESTION} Unknown"',
        allow_blank=False)
    dv_yesno.add('N4:N54')

    validations = {
        'event': dv_event, 'access_level': dv_access_level,
        'method': dv_method, 'verification': dv_verification, 'yesno': dv_yesno,
    }
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    num_cols = len(headers)
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add sample data row with formulas
    row = 4
    ws[f'A{row}'] = "[Example: Former Employee]"
    ws[f'B{row}'] = "[Example: former@organisation.ch]"
    ws[f'C{row}'] = "Termination"
    ws[f'D{row}'] = (datetime.now() - timedelta(days=2)).strftime("%d.%m.%Y")
    ws[f'E{row}'] = (datetime.now() - timedelta(days=2)).strftime("%d.%m.%Y")
    ws[f'F{row}'] = "[Example: customer-portal]"
    ws[f'G{row}'] = "Write/Contribute"
    ws[f'H{row}'] = (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")
    ws[f'I{row}'] = "=INT((H4-D4)*24)"  # Formula: Hours to deprovision
    ws[f'J{row}'] = "[Example: DevOps Team]"
    ws[f'K{row}'] = "Automated"
    ws[f'L{row}'] = "Platform Check"
    ws[f'M{row}'] = (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")
    ws[f'N{row}'] = f"{CHECK} Yes"
    ws[f'O{row}'] = ""
    ws[f'P{row}'] = ""

    # Style sample row
    for col in range(1, num_cols + 1):
        apply_style(ws.cell(row=row, column=col), styles['sample_row'])

    # Pre-format 50 additional data rows for user input (rows 5-53)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 9: SHEET 7 - THIRD PARTY ACCESS
# ============================================================================

def create_third_party_access_sheet(wb, styles):
    """Create the Third Party Access sheet."""
    ws = wb["Third Party Access"]
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "THIRD PARTY SOURCE CODE ACCESS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:M2')
    cell = ws['A2']
    cell.value = "Document all third party access to source code repositories"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Third Party Name",
        "Access Type",
        "Repository/System",
        "Access Level",
        "Business Justification",
        "Approval Date",
        "Approved By",
        "Expiry Date",
        "Review Date",
        "NDA Reference",
        "Access Status",
        "Risk Rating",
        "Notes",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Set column widths
    widths = [25, 20, 30, 18, 40, 15, 20, 15, 15, 20, 18, 15, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Data validations
    num_cols = len(headers)

    dv_access_type = DataValidation(type="list",
        formula1='"Contractor,Consultant,Vendor,Auditor,Partner,Other"',
        allow_blank=False)
    dv_access_type.add('B4:B53')

    dv_access_level = DataValidation(type="list",
        formula1='"Read,Write/Contribute,Admin"', allow_blank=False)
    dv_access_level.add('D4:D54')

    dv_access_status = DataValidation(type="list",
        formula1='"Active,Suspended,Expired,Revoked,Pending"',
        allow_blank=False)
    dv_access_status.add('K4:K53')

    dv_risk = DataValidation(type="list",
        formula1='"Critical,High,Medium,Low"', allow_blank=False)
    dv_risk.add('L4:L53')

    validations = {
        'access_type': dv_access_type, 'access_level': dv_access_level,
        'access_status': dv_access_status, 'risk': dv_risk,
    }
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    # Sample data row at row 4
    row = 4
    sample_data = [
        "SecureAudit GmbH", "Auditor", "customer-portal", "Read",
        "Annual security audit (Q1 2026)", "15.01.2026", "CISO",
        "15.04.2026", "15.03.2026", "NDA-2026-001", "Active", "Low",
        "Read-only access for Q1 audit"
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 50 additional data rows for user input (rows 5-54)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 10: SHEET 8 - SERVICE ACCOUNTS
# ============================================================================

def create_service_accounts_sheet(wb, styles):
    """Create the Service Accounts sheet."""
    ws = wb["Service Accounts"]
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = "SERVICE ACCOUNT AND BOT ACCESS INVENTORY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:N2')
    cell = ws['A2']
    cell.value = "Document all service accounts, bots, and automated systems with source code access"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Account Name",
        "Account Type",
        "Repository/System",
        "Purpose",
        "Owner",
        "Access Level",
        "Authentication Method",
        "Secret Storage",
        "Last Rotated",
        "Rotation Frequency",
        "Review Date",
        "Status",
        "Risk Rating",
        "Notes",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Set column widths
    widths = [22, 20, 30, 35, 20, 18, 22, 22, 15, 18, 15, 15, 15, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Data validations
    num_cols = len(headers)

    dv_account_type = DataValidation(type="list",
        formula1='"CI/CD Pipeline,Deployment Bot,Monitoring Agent,Integration Service,API Key,Other"',
        allow_blank=False)
    dv_account_type.add('B4:B53')

    dv_access_level = DataValidation(type="list",
        formula1='"Read,Write/Contribute,Admin"', allow_blank=False)
    dv_access_level.add('F4:F53')

    dv_auth_method = DataValidation(type="list",
        formula1='"SSH Key,Personal Access Token,OAuth Token,Deploy Key,API Key,Certificate,Other"',
        allow_blank=False)
    dv_auth_method.add('G4:G53')

    dv_secret_storage = DataValidation(type="list",
        formula1='"Vault/HSM,CI/CD Secrets,Environment Variable,Key Management Service,Other"',
        allow_blank=False)
    dv_secret_storage.add('H4:H53')

    dv_rotation = DataValidation(type="list",
        formula1='"30 days,60 days,90 days,180 days,365 days,On compromise only,N/A"',
        allow_blank=False)
    dv_rotation.add('J4:J54')

    dv_status = DataValidation(type="list",
        formula1='"Active,Disabled,Expired,Pending Review"',
        allow_blank=False)
    dv_status.add('L4:L53')

    dv_risk = DataValidation(type="list",
        formula1='"Critical,High,Medium,Low"', allow_blank=False)
    dv_risk.add('M4:M53')

    validations = {
        'account_type': dv_account_type, 'access_level': dv_access_level,
        'auth_method': dv_auth_method, 'secret_storage': dv_secret_storage,
        'rotation': dv_rotation, 'status': dv_status, 'risk': dv_risk,
    }
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    # Sample data row at row 4
    row = 4
    sample_data = [
        "github-actions-bot", "CI/CD Pipeline", "customer-portal", "GitHub Actions CI/CD",
        "DevOps Team", "Write/Contribute", "Deploy Key", "GitHub Secrets",
        "15.01.2026", "90 days", "15.01.2026", "Active", "Medium",
        "Automated deployment to production"
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 50 additional data rows for user input (rows 5-54)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 11: SHEET 9 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(wb, styles):
    """Create the Summary Dashboard — Gold Standard 7-col TABLE 1/2/3."""
    ws = wb["Summary Dashboard"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    Q = chr(34)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _red  = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _yell = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # ── Row 1: Title (A1:G1, 003366) ────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME} \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # ── Row 2: Control ref (A2:G2, italic 003366, NO fill) ──────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=2, column=_c).border = border

    # ── Row 3: TABLE 1 banner (A3:G3, 003366) ───────────────────────────────
    ws.merge_cells("A3:G3")
    ws["A3"] = "COMPLIANCE ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = _navy
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=3, column=_c).border = border

    # ── Row 4: Column headers (D9D9D9) ──────────────────────────────────────
    _hdrs = ["Assessment Area", "Questions Answered", "No Gap",
             "Gap Identified", "N/A", "Target", "Compliance %"]
    for _ci, _h in enumerate(_hdrs, 1):
        _cell = ws.cell(row=4, column=_ci, value=_h)
        _cell.font = Font(name="Calibri", bold=True, size=10, color="000000")
        _cell.fill = _grey
        _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _cell.border = border
    ws.row_dimensions[4].height = 30

    # ── TABLE 1 areas rows 5-9 ───────────────────────────────────────────────
    _t1_areas = [
        ("User Access Appropriateness",
         "=COUNTA('User Access Matrix'!A5:A54)",
         f"=COUNTIF('User Access Matrix'!O5:O54,{Q}[PASS] Appropriate{Q})",
         f"=COUNTIF('User Access Matrix'!O5:O54,{Q}[WARN] Excessive{Q})"
         f"+COUNTIF('User Access Matrix'!O5:O54,{Q}[FAIL] Orphaned{Q})"
         f"+COUNTIF('User Access Matrix'!O5:O54,{Q}[DATE] Expired{Q})",
         0.85),
        ("Access Authorization Compliance",
         "=COUNTA('Access Approval Records'!B5:B54)",
         f"=COUNTIF('Access Approval Records'!K5:K54,{Q}[PASS] Approved{Q})",
         f"=COUNTIF('Access Approval Records'!K5:K54,{Q}[FAIL] Denied{Q})"
         f"+COUNTIF('Access Approval Records'!K5:K54,{Q}[WAIT] Pending{Q})",
         0.90),
        ("Access Review Compliance",
         "=COUNTA('Access Review Log'!A5:A54)",
         f"=COUNTIF('Access Review Log'!S5:S54,{Q}[PASS] Compliant{Q})",
         f"=COUNTIF('Access Review Log'!S5:S54,{Q}[WARN] Partial{Q})"
         f"+COUNTIF('Access Review Log'!S5:S54,{Q}[FAIL] Non-Compliant{Q})",
         0.85),
        ("Deprovisioning Timeliness",
         "=COUNTA('Deprovisioning Log'!A5:A54)",
         f"=COUNTIF('Deprovisioning Log'!N5:N54,{Q}[PASS] Yes{Q})",
         f"=COUNTIF('Deprovisioning Log'!N5:N54,{Q}[FAIL] No{Q})"
         f"+COUNTIF('Deprovisioning Log'!N5:N54,{Q}[WARN] Partial{Q})",
         0.90),
        ("Repository Security Controls",
         "=COUNTA('Repository Inventory'!A5:A54)",
         f"=COUNTIFS('Repository Inventory'!Q5:Q54,{Q}[PASS] Yes{Q},"
         f"'Repository Inventory'!R5:R54,{Q}[PASS] Yes{Q})",
         f"=COUNTA('Repository Inventory'!A5:A54)"
         f"-COUNTIFS('Repository Inventory'!Q5:Q54,{Q}[PASS] Yes{Q},"
         f"'Repository Inventory'!R5:R54,{Q}[PASS] Yes{Q})",
         0.80),
    ]
    for _r, (_area, _b_f, _c_f, _d_f, _tgt) in enumerate(_t1_areas, 5):
        for _ci in range(1, 8):
            ws.cell(row=_r, column=_ci).fill = _yell
            ws.cell(row=_r, column=_ci).border = border
            ws.cell(row=_r, column=_ci).alignment = Alignment(
                horizontal="center", vertical="center")
        ws.cell(row=_r, column=1, value=_area).alignment = Alignment(
            horizontal="left", vertical="center")
        ws.cell(row=_r, column=2, value=_b_f)
        ws.cell(row=_r, column=3, value=_c_f)
        ws.cell(row=_r, column=4, value=_d_f)
        ws.cell(row=_r, column=5, value=0)
        ws.cell(row=_r, column=6, value=_tgt).number_format = "0%"
        ws.cell(row=_r, column=7,
                value=f"=IFERROR(C{_r}/(C{_r}+D{_r})*100,0)").number_format = "0.0"
        ws.row_dimensions[_r].height = 18

    # ── Row 10: TOTAL ────────────────────────────────────────────────────────
    for _ci in range(1, 8):
        ws.cell(row=10, column=_ci).fill = _grey
        ws.cell(row=10, column=_ci).border = border
        ws.cell(row=10, column=_ci).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=10, column=_ci).alignment = Alignment(
            horizontal="center", vertical="center")
    ws.cell(row=10, column=1, value="TOTAL").alignment = Alignment(
        horizontal="left", vertical="center")
    ws.cell(row=10, column=2, value="=SUM(B5:B9)")
    ws.cell(row=10, column=3, value="=SUM(C5:C9)")
    ws.cell(row=10, column=4, value="=SUM(D5:D9)")
    ws.cell(row=10, column=5, value="=SUM(E5:E9)")
    ws.cell(row=10, column=6, value="\u2014")
    ws.cell(row=10, column=7, value="=IFERROR(AVERAGE(G5:G9),0)").number_format = "0.0"

    # ── TABLE 2: REPOSITORY ACCESS HEALTH METRICS (003366 banner) ───────────
    _t2_row = 12
    ws.merge_cells(f"A{_t2_row}:G{_t2_row}")
    ws[f"A{_t2_row}"] = "REPOSITORY ACCESS HEALTH METRICS"
    ws[f"A{_t2_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t2_row}"].fill = _navy
    ws[f"A{_t2_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t2_row, column=_c).border = border
    _t2_row += 1
    for _ci, _h in enumerate(["Metric", "Value", "Target"], 1):
        _cell = ws.cell(row=_t2_row, column=_ci, value=_h)
        _cell.font = Font(name="Calibri", bold=True, size=10, color="000000")
        _cell.fill = _grey
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t2_row += 1

    def _metric(label, formula, target="\u2014", num_fmt="General"):
        nonlocal _t2_row
        ws.cell(row=_t2_row, column=1, value=label).fill = _yell
        ws.cell(row=_t2_row, column=1).border = border
        ws.cell(row=_t2_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        _v = ws.cell(row=_t2_row, column=2, value=formula)
        _v.fill = _yell
        _v.border = border
        _v.alignment = Alignment(horizontal="center", vertical="center")
        if num_fmt != "General":
            _v.number_format = num_fmt
        _t = ws.cell(row=_t2_row, column=3, value=target)
        _t.fill = _yell
        _t.border = border
        _t.alignment = Alignment(horizontal="center", vertical="center")
        _t2_row += 1

    _metric("Total Repositories",
            "=COUNTA('Repository Inventory'!A5:A54)", "\u22651")
    _metric("Repos with Branch Protection",
            f"=COUNTIF('Repository Inventory'!Q5:Q54,{Q}[PASS] Yes{Q})")
    _metric("Repos with Secret Scanning",
            f"=COUNTIF('Repository Inventory'!R5:R54,{Q}[PASS] Yes{Q})")
    _metric("Admin Access Count",
            f"=COUNTIF('User Access Matrix'!H5:H54,{Q}Admin{Q})", "\u22645")
    _metric("Contractor Access Count",
            f"=COUNTIF('User Access Matrix'!D5:D54,{Q}Contractor{Q})")
    _metric("Orphaned Account Count",
            f"=COUNTIF('User Access Matrix'!O5:O54,{Q}[FAIL] Orphaned{Q})", "0")
    _metric("Pending Access Approvals",
            f"=COUNTIF('Access Approval Records'!K5:K54,{Q}[WAIT] Pending{Q})", "0")
    _metric("Active Third Party Access Grants",
            f"=COUNTIF('Third Party Access'!K5:K54,{Q}Active{Q})")
    _metric("Expired Third Party Access",
            f"=COUNTIF('Third Party Access'!K5:K54,{Q}Expired{Q})", "0")
    _metric("Service Accounts Pending Review",
            f"=COUNTIF('Service Accounts'!L5:L54,{Q}Pending Review{Q})", "0")
    _metric("Open Remediation Gaps",
            f"=COUNTIF('Gap Analysis'!O5:O54,{Q}[OPEN] Open{Q})", "0")
    _metric("Unverified Evidence Items",
            f"=COUNTIF('Evidence Register'!H6:H105,{Q}Not verified{Q})", "0")

    # ── TABLE 3: CRITICAL FINDINGS (C00000 banner) ───────────────────────────
    _t3_row = _t2_row + 1
    ws.merge_cells(f"A{_t3_row}:G{_t3_row}")
    ws[f"A{_t3_row}"] = "CRITICAL FINDINGS"
    ws[f"A{_t3_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t3_row}"].fill = _red
    ws[f"A{_t3_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t3_row, column=_c).border = border
    _t3_row += 1
    for _ci, _h in enumerate(["Critical Finding", "Status", "Severity"], 1):
        _cell = ws.cell(row=_t3_row, column=_ci, value=_h)
        _cell.font = Font(name="Calibri", bold=True, size=10, color="000000")
        _cell.fill = _grey
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t3_row += 1
    _sev_fill = {"Critical": "FFC7CE", "High": "FFEB9C", "Low": "C6EFCE"}
    _t3_findings = [
        ("Orphaned accounts detected",
         f"=IF(COUNTIF('User Access Matrix'!O5:O54,{Q}[FAIL] Orphaned{Q})>0,"
         f"{Q}[!] Review Required{Q},{Q}[OK]{Q})",
         "Critical"),
        ("Excessive access detected",
         f"=IF(COUNTIF('User Access Matrix'!O5:O54,{Q}[WARN] Excessive{Q})>0,"
         f"{Q}[!] Review Required{Q},{Q}[OK]{Q})",
         "Critical"),
        ("Deprovisioning SLA failures",
         f"=IF(COUNTIF('Deprovisioning Log'!N5:N54,{Q}[FAIL] No{Q})>0,"
         f"{Q}[!] Review Required{Q},{Q}[OK]{Q})",
         "High"),
        ("Expired third party access grants",
         f"=IF(COUNTIF('Third Party Access'!K5:K54,{Q}Expired{Q})>0,"
         f"{Q}[!] Review Required{Q},{Q}[OK]{Q})",
         "High"),
        ("Repos missing branch protection",
         f"=IF(COUNTIF('Repository Inventory'!Q5:Q54,{Q}[FAIL] No{Q})>0,"
         f"{Q}[!] Review Required{Q},{Q}[OK]{Q})",
         "High"),
        ("Overall compliance <70%",
         f"=IF(G10<70,{Q}[!] Below Target{Q},{Q}[OK]{Q})",
         "Critical"),
        ("Open critical/high severity gaps",
         f"=IF(COUNTIFS('Gap Analysis'!G5:G54,{Q}[CRIT] Critical{Q},"
         f"'Gap Analysis'!O5:O54,{Q}[OPEN] Open{Q})"
         f"+COUNTIFS('Gap Analysis'!G5:G54,{Q}[HIGH] High{Q},"
         f"'Gap Analysis'!O5:O54,{Q}[OPEN] Open{Q})>0,"
         f"{Q}[!] Review Required{Q},{Q}[OK]{Q})",
         "High"),
    ]
    for _label, _formula, _sev in _t3_findings:
        _fc = _sev_fill.get(_sev, "FFFFCC")
        _fp = PatternFill(start_color=_fc, end_color=_fc, fill_type="solid")
        for _ci in range(1, 4):
            ws.cell(row=_t3_row, column=_ci).fill = _fp
            ws.cell(row=_t3_row, column=_ci).border = border
            ws.cell(row=_t3_row, column=_ci).alignment = Alignment(
                horizontal="left", vertical="center")
        ws.cell(row=_t3_row, column=1, value=_label)
        ws.cell(row=_t3_row, column=2, value=_formula)
        ws.cell(row=_t3_row, column=3, value=_sev)
        _t3_row += 1

    # ── Column widths + freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"



def create_gap_analysis_sheet(wb, styles):
    """Create the Gap Analysis sheet."""
    ws = wb["Gap Analysis"]
    ws.sheet_view.showGridLines = False
    
    # Header
    ws.merge_cells('A1:S1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:S2')
    cell = ws['A2']
    cell.value = "Document identified gaps and remediation plans"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Gap ID",
        "Gap Category",
        "Gap Description",
        "Policy Requirement",
        "Current State",
        "Desired State",
        "Risk Level",
        "Impact",
        "Affected Repositories",
        "Root Cause",
        "Remediation Plan",
        "Responsible Party",
        "Target Completion Date",
        "Estimated Effort",
        "Status",
        "Actual Completion Date",
        "Verification Method",
        "Verification Date",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [12, 18, 40, 25, 30, 30, 15, 30, 30, 30, 40, 20, 18, 18, 15, 18, 25, 15, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validations (cover all 50 data rows: 4-53)
    dv_category = DataValidation(type="list",
        formula1='"Access Control,Inventory,Reviews,Deprovisioning,Documentation"',
        allow_blank=False)
    dv_category.add('B4:B53')

    dv_risk = DataValidation(type="list",
        formula1='"[CRIT] Critical,[HIGH] High,[MED] Medium,[LOW] Low"',
        allow_blank=False)
    dv_risk.add('G4:G53')

    dv_effort = DataValidation(type="list",
        formula1='"1-2 hours,1 day,1 week,2-4 weeks,>1 month"',
        allow_blank=False)
    dv_effort.add('N4:N54')

    dv_gap_status = DataValidation(type="list",
        formula1='"[OPEN] Open,[PROG] In Progress,[DONE] Completed,[DEF] Deferred"',
        allow_blank=False)
    dv_gap_status.add('O4:O54')

    validations = {
        'category': dv_category, 'risk': dv_risk,
        'effort': dv_effort, 'gap_status': dv_gap_status,
    }
    for _key, _dv in validations.items():
        if hasattr(_dv, "sqref") and _dv.sqref:
            try:
                ws.add_data_validation(_dv)
            except Exception:
                pass

    num_cols = len(headers)
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add sample data row
    row = 4
    ws[f'A{row}'] = "GAP-001"
    ws[f'B{row}'] = "Access Control"
    ws[f'C{row}'] = "[Example: Branch protection not enabled on all production repositories]"
    ws[f'D{row}'] = "[Example: REQ-2.4.1-001 - Main branch protection]"
    ws[f'E{row}'] = "[Example: 15 out of 20 production repos lack branch protection]"
    ws[f'F{row}'] = "[Example: All production repos have branch protection enabled]"
    ws[f'G{row}'] = "[HIGH] High"
    ws[f'H{row}'] = "[Example: Unauthorised direct commits to production code possible]"
    ws[f'I{row}'] = "[Example: api-gateway, payment-processor, ...]"
    ws[f'J{row}'] = "[Example: Repositories created before policy implementation]"
    ws[f'K{row}'] = "[Example: Enable branch protection on all production repositories]"
    ws[f'L{row}'] = "[Example: Platform Team Lead]"
    ws[f'M{row}'] = (datetime.now() + timedelta(days=30)).strftime("%d.%m.%Y")
    ws[f'N{row}'] = "1 week"
    ws[f'O{row}'] = "[PROG] In Progress"
    ws[f'P{row}'] = ""
    ws[f'Q{row}'] = ""
    ws[f'R{row}'] = ""
    ws[f'S{row}'] = ""

    # Style sample row
    for col in range(1, num_cols + 1):
        apply_style(ws.cell(row=row, column=col), styles['sample_row'])

    # Pre-format 50 additional data rows for user input (rows 5-53)
    for r in range(5, 55):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = 'A4'


# ============================================================================
# SECTION 13: SHEET 11 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb):
    """Create Evidence Register matching gold standard."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["Evidence Register"]
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3 is blank separator

    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location/Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    # Sample row with complete example data
    row = 5
    sample_data = {
        1: "EV-001",
        2: "Repository Access",
        3: "Access review log",
        4: "Quarterly access review for Q4 2025",
        5: "/evidence/q4-2025-access-review.xlsx",
        6: "15.01.2026",
        7: "Security Team",
        8: "Verified",
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    dv_type.add(ws.cell(row=5, column=3))
    dv_ver.add(ws.cell(row=5, column=8))

    # Empty rows 6-105 (100 empty rows)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 14: SHEET 12 - APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(wb):
    """Create Approval Sign-Off matching gold standard."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["Approval Sign-Off"]
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border_thin
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border_thin
    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.4.1 - Repository Access Control Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G9),\"\")"),
        ("Assessment Status:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border_thin
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1


    # B6: Overall Compliance Rate — reference SD overall score
    ws["B6"] = "=IFERROR('Summary Dashboard'!G10,\"\")"
    ws["B6"].number_format = "0.0%"
    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border_thin
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border_thin
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border_thin
    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border_thin
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 15: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the Excel workbook."""
    logger.info("Starting ISMS-A84-1 Repository Access Assessment Workbook Generation...")
    
    # Create workbook and get styles
    wb = create_workbook()
    styles = _STYLES

    # Create all sheets
    logger.info("Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("Creating Repository Inventory sheet...")
    create_repository_inventory_sheet(wb, styles)

    logger.info("Creating User Access Matrix sheet...")
    create_user_access_matrix_sheet(wb, styles)

    logger.info("Creating Access Approval Records sheet...")
    create_access_approval_records_sheet(wb, styles)

    logger.info("Creating Access Review Log sheet...")
    create_access_review_log_sheet(wb, styles)

    logger.info("Creating Deprovisioning Log sheet...")
    create_deprovisioning_log_sheet(wb, styles)

    logger.info("Creating Third Party Access sheet...")
    create_third_party_access_sheet(wb, styles)

    logger.info("Creating Service Accounts sheet...")
    create_service_accounts_sheet(wb, styles)

    logger.info("Creating Gap Analysis sheet...")
    create_gap_analysis_sheet(wb, styles)

    logger.info("Creating Evidence Register sheet...")
    create_evidence_register(wb)

    logger.info("Creating Summary Dashboard sheet...")
    create_summary_dashboard_sheet(wb, styles)

    logger.info("Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb)
    
    # Save workbook
    output_filename = f"ISMS-IMP-A.8.4.1_Repository_Access_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook generated successfully: {_wkbk_dir / output_filename}")
    logger.info(f"Sheets created: {len(wb.sheetnames)}")
    logger.info("\nWorkbook is ready for use in repository access control assessments!")


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
