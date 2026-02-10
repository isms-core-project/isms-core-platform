#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.4.3 - Source Code Security Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.4: Access to Source Code
Assessment Domain 3 of 3: Executive Compliance Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dashboard requirements, KPI definitions, and
reporting structure.

Key customization areas:
1. External workbook paths (update to your actual file locations)
2. KPI definitions and thresholds (per your governance requirements)
3. Compliance scoring weights (aligned with your risk framework)
4. Repository criticality tiers (based on your classification)
5. Approval workflow (per your organizational structure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for source code)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Consolidates data from repository access and branch protection assessments
into unified executive dashboard for ISO 27001:2022 Control A.8.4 compliance
oversight.

**Dashboard Consolidates:**
- ISMS-IMP-A.8.4.1: Repository Access Control Assessment
- ISMS-IMP-A.8.4.2: Branch Protection Assessment

**Generated Dashboard Structure:**
1. Instructions & Legend - Dashboard usage guidance
2. Executive Summary - High-level compliance status
3. Repository Overview - Access control compliance summary
4. Branch Protection Status - Protection rule compliance
5. Access Metrics - Permission and access KPIs
6. Risk Analysis - Repository risk distribution
7. Trend Analysis - Historical compliance tracking
8. Gap Remediation - Priority remediation tracking
9. Evidence Summary - Audit evidence consolidation
10. Action Items - Open tasks and owners
11. Approval & Sign-Off - Executive approval workflow

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a84_3_compliance_dashboard.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-A84-3-Source-Code-Security-Dashboard.xlsx
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)







# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.4.3"
WORKBOOK_NAME = "Source Code Security Compliance Dashboard"
CONTROL_ID = "A.8.4"
CONTROL_NAME = "Access to Source Code"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # f"{CHECK}" Green checkmark
XMARK = '\u274C'      # f"{XMARK}" Red X
WARNING = '\u26A0'    # f"{WARNING}"  Warning sign
HOURGLASS = '\u23F3'  # f"{HOURGLASS}" Hourglass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow
QUESTION = '\u2753'   # f"{QUESTION}" Question mark
CALENDAR = '\u1F4C5' # f"{CALENDAR}" Calendar
CHECKMARK = '\u2714'  # f"{CHECKMARK}"  Check mark
MINUS = '\u2796'      # f"{MINUS}" Minus
CHART = '\U0001F4CA' # f"{CHART}" Chart
TARGET = '\U0001F3AF' # f"{TARGET}" Target

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLES
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Executive_Summary",
        "Repository_Overview",
        "Access_Control_Metrics",
        "Branch_Protection_Metrics",
        "Secret_Management_Metrics",
        "Third_Party_Access",
        "Trend_Analysis",
        "Gap_Priority_Matrix",
        "Action_Items",
        "Evidence_Summary",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "kpi_label": {
            "font": Font(name="Calibri", size=11, bold=True),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "kpi_value": {
            "font": Font(name="Calibri", size=14, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "score_excellent": {
            "font": Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "score_good": {
            "font": Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "score_fair": {
            "font": Font(name="Calibri", size=18, bold=True),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "score_poor": {
            "font": Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text if hasattr(style_dict["alignment"], 'wrap_text') else False
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: SHEET 1 - EXECUTIVE SUMMARY
# ============================================================================

def create_executive_summary_sheet(wb, styles):
    """Create Executive_Summary dashboard."""
    ws = wb["Executive_Summary"]
    
    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "SOURCE CODE SECURITY COMPLIANCE DASHBOARD"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 50
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = f"Executive Summary - {datetime.now().strftime('%B %Y')}"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 30
    
    # Overall Compliance Score (Large Display)
    row = 4
    ws.merge_cells(f'A{row}:B{row+2}')
    cell = ws[f'A{row}']
    cell.value = "OVERALL\nCOMPLIANCE\nSCORE"
    cell.font = Font(name="Calibri", size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws.merge_cells(f'C{row}:D{row+2}')
    cell = ws[f'C{row}']
    cell.value = "92%"  # Example score - to be calculated from assessment data
    apply_style(cell, styles['score_excellent'])
    ws.row_dimensions[row].height = 60
    
    ws.merge_cells(f'E{row}:F{row+2}')
    cell = ws[f'E{row}']
    cell.value = "Target: ≥85%"
    cell.font = Font(name="Calibri", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # KPI Summary
    row += 4
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "KEY PERFORMANCE INDICATORS"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    kpi_headers = ["Metric", "Current", "Target", "Status", "Weight", "Trend"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    kpis = [
        ("Repository Access Control", "92%", "≥90%", "🟢 Low Risk", "35%", "↗ Improving"),
        ("Branch Protection", "88%", "≥85%", "🟢 Low Risk", "35%", "→ Stable"),
        ("Secret Management", "78%", "≥80%", "🟡 Medium Risk", "20%", "↗ Improving"),
        ("Third-Party Access", "95%", "≥90%", "🟢 Low Risk", "10%", "→ Stable"),
    ]
    
    for metric, current, target, status, weight, trend in kpis:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = current
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status
        ws[f'E{row}'] = weight
        ws[f'F{row}'] = trend
        row += 1
    
    # Critical Findings
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "CRITICAL FINDINGS SUMMARY"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    findings_headers = ["Severity", "Count", "Category", "Avg Age (Days)", "Status"]
    for col_idx, header in enumerate(findings_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    findings = [
        ("🔴 Critical", "2", "Secret Exposure", "3", "🟡 In Remediation"),
        ("🟠 High", "5", "Orphaned Access", "12", "🟢 Remediation Complete"),
        ("🟡 Medium", "8", "Branch Protection", "25", "🟡 In Progress"),
    ]
    
    for severity, count, category, age, status in findings:
        ws[f'A{row}'] = severity
        ws[f'B{row}'] = count
        ws[f'C{row}'] = category
        ws[f'D{row}'] = age
        ws[f'E{row}'] = status
        row += 1
    
    # Action Items Summary
    row += 2
    ws[f'A{row}'] = "TOP 5 ACTION ITEMS"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    row += 1
    
    actions = [
        "1. Remediate 2 critical secret exposures (Target: 2 days)",
        "2. Remove 5 orphaned accounts across platforms (Target: 7 days)",
        "3. Enable branch protection on 8 production repositories (Target: 14 days)",
        "4. Complete quarterly access review for 12 overdue repositories (Target: 7 days)",
        "5. Update contractor access documentation for 3 expired contracts (Target: 3 days)",
    ]
    
    for action in actions:
        ws[f'A{row}'] = action
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15


# ============================================================================
# SECTION 3: SHEET 2 - REPOSITORY OVERVIEW
# ============================================================================

def create_repository_overview_sheet(wb, styles):
    """Create Repository_Overview sheet."""
    ws = wb["Repository_Overview"]
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "REPOSITORY OVERVIEW"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 40
    
    row = 3
    ws[f'A{row}'] = "REPOSITORY STATISTICS"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    row += 1
    stats = [
        ("Total Repositories", "125"),
        ("Production Code", "45 (36%)"),
        ("Internal Tools", "60 (48%)"),
        ("Open Source", "15 (12%)"),
        ("Archived", "5 (4%)"),
        ("", ""),
        ("Repositories by Platform", ""),
        ("GitHub", "80 (64%)"),
        ("GitLab", "30 (24%)"),
        ("Bitbucket", "10 (8%)"),
        ("Azure DevOps", "5 (4%)"),
    ]
    
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if label and not value:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 2
    ws[f'A{row}'] = "SECURITY STATUS SUMMARY"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    row += 1
    security = [
        ("Access Control Enabled", "122/125 (98%)", f"{CHECK}"),
        ("Branch Protection Configured", "40/45 (89%)", f"{WARNING}"),
        ("Secret Scanning Active", "125/125 (100%)", f"{CHECK}"),
        ("Active Secret Findings", "5 total", "🔴"),
        ("Quarterly Review Complete", "120/125 (96%)", f"{CHECK}"),
    ]
    
    for metric, value, status in security:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'C{row}'] = status
        ws[f'C{row}'].alignment = Alignment(horizontal="center")
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 8


# ============================================================================
# SECTION 4: SHEET 3 - ACCESS CONTROL METRICS
# ============================================================================

def create_access_control_metrics_sheet(wb, styles):
    """Create Access_Control_Metrics sheet."""
    ws = wb["Access_Control_Metrics"]
    
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "ACCESS CONTROL METRICS"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["Metric", "Current Value", "Target", "Status", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    metrics = [
        ("Repository Inventory Completeness", "98%", "100%", f"{WARNING} Partial", "2 shadow repos detected"),
        ("Appropriate Access Rate", "92%", "≥95%", f"{WARNING} Partial", "10 excessive access instances"),
        ("Orphaned Account Rate", "0.4%", "0%", f"{WARNING} Partial", "5 orphaned accounts"),
        ("Access Review Completion", "96%", "100%", f"{WARNING} Partial", "5 overdue reviews"),
        ("Deprovisioning SLA Compliance", "98%", "≥95%", f"{CHECK} Compliant", "Avg 8 hours"),
        ("Service Account Documentation", "100%", "100%", f"{CHECK} Compliant", "All documented"),
        ("MFA Adoption", "100%", "100%", f"{CHECK} Compliant", "All users enabled"),
    ]
    
    for metric, current, target, status, notes in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = current
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status
        ws[f'E{row}'] = notes
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40


# ============================================================================
# SECTION 5: SHEET 4 - BRANCH PROTECTION METRICS
# ============================================================================

def create_branch_protection_metrics_sheet(wb, styles):
    """Create Branch_Protection_Metrics sheet."""
    ws = wb["Branch_Protection_Metrics"]
    
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "BRANCH PROTECTION METRICS"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["Metric", "Current Value", "Target", "Status", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    metrics = [
        ("Production Repos Protected", "40/45 (89%)", "100%", f"{WARNING} Partial", "5 repos missing protection"),
        ("PR Enforcement Rate", "97%", "≥95%", f"{CHECK} Compliant", "12 direct commits last month"),
        ("Avg Reviewers per PR", "2.1", "≥2.0", f"{CHECK} Compliant", "Production code standard"),
        ("Status Check Coverage", "42/45 (93%)", "100%", f"{WARNING} Partial", "3 repos without CI/CD"),
        ("Signed Commit Adoption", "78%", "≥80%", f"{WARNING} Partial", "Improving from 65% last quarter"),
        ("Self-Approval Incidents", "0", "0", f"{CHECK} Compliant", "No violations"),
        ("Protection Rule Exceptions", "2 active", "Minimize", f"{CHECK} Compliant", "Both justified and time-bound"),
    ]
    
    for metric, current, target, status, notes in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = current
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status
        ws[f'E{row}'] = notes
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40


# ============================================================================
# SECTION 6: SHEET 5 - SECRET MANAGEMENT METRICS
# ============================================================================

def create_secret_management_metrics_sheet(wb, styles):
    """Create Secret_Management_Metrics sheet."""
    ws = wb["Secret_Management_Metrics"]
    
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "SECRET MANAGEMENT METRICS"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["Metric", "Current Value", "Target", "Status", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    metrics = [
        ("Scan Coverage", "125/125 (100%)", "100%", f"{CHECK} Compliant", "All repos scanned daily"),
        ("Active Secret Findings", "5 total", "0", "🔴 Critical", "2 critical, 3 medium"),
        ("Critical Secret MTTR", "4 hours", "<4 hours", f"{CHECK} Compliant", "Within SLA"),
        ("Medium Secret MTTR", "18 hours", "<24 hours", f"{CHECK} Compliant", "Within SLA"),
        ("Secrets Rotated After Exposure", "100%", "100%", f"{CHECK} Compliant", "All rotated immediately"),
        ("Pre-Commit Hook Adoption", "85%", "≥90%", f"{WARNING} Partial", "Improving from 70%"),
        ("Developer Training Completion", "95%", "100%", f"{WARNING} Partial", "Q1 training scheduled"),
        ("False Positive Rate", "8%", "<10%", f"{CHECK} Compliant", "Tuning ongoing"),
    ]
    
    for metric, current, target, status, notes in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = current
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status
        ws[f'E{row}'] = notes
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40


# ============================================================================
# SECTION 7: REMAINING SHEETS (Simplified)
# ============================================================================

def create_third_party_access_sheet(wb, styles):
    """Create Third_Party_Access sheet."""
    ws = wb["Third_Party_Access"]
    
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "THIRD-PARTY ACCESS METRICS"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["Metric", "Current Value", "Target", "Status", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    metrics = [
        ("Active Contractors", "12", "N/A", "ℹ️ Info", "Across 8 repositories"),
        ("Time-Bound Access Compliance", "100%", "100%", f"{CHECK} Compliant", "All have end dates"),
        ("Expired Access Removal", "100%", "100%", f"{CHECK} Compliant", "Auto-expiration enabled"),
        ("Enhanced Review (2+ Reviewers)", "100%", "100%", f"{CHECK} Compliant", "All contractor PRs"),
        ("NDA Documentation", "100%", "100%", f"{CHECK} Compliant", "All contractors signed"),
        ("IP Assignment Documentation", "100%", "100%", f"{CHECK} Compliant", "All documented"),
    ]
    
    for metric, current, target, status, notes in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = current
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status
        ws[f'E{row}'] = notes
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40


def create_trend_analysis_sheet(wb, styles):
    """Create Trend_Analysis sheet."""
    ws = wb["Trend_Analysis"]
    
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "12-MONTH COMPLIANCE TREND ANALYSIS"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["Month", "Overall Score", "Access Control", "Branch Protection", "Secret Mgmt", "Third-Party", "Trend"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Sample trend data
    row += 1
    base_date = datetime.now()
    for i in range(12, 0, -1):
        month = (base_date - timedelta(days=30*i)).strftime("%b %Y")
        ws[f'A{row}'] = month
        ws[f'B{row}'] = f"{82 + i}%"
        ws[f'C{row}'] = f"{85 + i}%"
        ws[f'D{row}'] = f"{80 + i}%"
        ws[f'E{row}'] = f"{75 + i}%"
        ws[f'F{row}'] = f"{90 + i//2}%"
        ws[f'G{row}'] = "↗" if i > 6 else "→"
        row += 1


def create_gap_priority_matrix_sheet(wb, styles):
    """Create Gap_Priority_Matrix sheet."""
    ws = wb["Gap_Priority_Matrix"]
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "GAP PRIORITY MATRIX"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["Priority", "Gap Description", "Impact", "Effort", "Target Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    gaps = [
        ("🔴 Critical", "2 critical secrets exposed", "High", "2 days", (datetime.now() + timedelta(days=2)).strftime("%d.%m.%Y"), "🟡 In Progress"),
        ("🟠 High", "5 repos without branch protection", "Medium", "2 weeks", (datetime.now() + timedelta(days=14)).strftime("%d.%m.%Y"), "🟡 In Progress"),
        ("🟡 Medium", "Pre-commit hooks adoption <90%", "Low", "1 month", (datetime.now() + timedelta(days=30)).strftime("%d.%m.%Y"), "🔴 Not Started"),
    ]
    
    for priority, gap, impact, effort, target, status in gaps:
        ws[f'A{row}'] = priority
        ws[f'B{row}'] = gap
        ws[f'C{row}'] = impact
        ws[f'D{row}'] = effort
        ws[f'E{row}'] = target
        ws[f'F{row}'] = status
        row += 1
    
    ws.column_dimensions['B'].width = 40


def create_action_items_sheet(wb, styles):
    """Create Action_Items sheet."""
    ws = wb["Action_Items"]
    
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "ACTION ITEMS TRACKER"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["ID", "Action Item", "Owner", "Priority", "Target Date", "Status", "Days Overdue"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    actions = [
        ("ACT-001", "Remediate secret exposure in api-gateway", "Security Team", "🔴 Critical", (datetime.now() + timedelta(days=1)).strftime("%d.%m.%Y"), "🟡 In Progress", "0"),
        ("ACT-002", "Enable branch protection on 5 repos", "Platform Team", "🟠 High", (datetime.now() + timedelta(days=14)).strftime("%d.%m.%Y"), "🟡 In Progress", "0"),
        ("ACT-003", "Complete overdue access reviews", "Repo Owners", "🟡 Medium", (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y"), "🔴 Not Started", "0"),
    ]
    
    for action_id, action, owner, priority, target, status, overdue in actions:
        ws[f'A{row}'] = action_id
        ws[f'B{row}'] = action
        ws[f'C{row}'] = owner
        ws[f'D{row}'] = priority
        ws[f'E{row}'] = target
        ws[f'F{row}'] = status
        ws[f'G{row}'] = overdue
        row += 1
    
    ws.column_dimensions['B'].width = 45


def create_evidence_summary_sheet(wb, styles):
    """Create Evidence_Summary sheet."""
    ws = wb["Evidence_Summary"]
    
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "EVIDENCE SUMMARY - AUDIT READINESS"
    apply_style(cell, styles['header'])
    
    row = 3
    headers = ["Evidence Type", "Required", "Collected", "Completeness"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    evidence = [
        ("Repository Access Reports", "Quarterly", "Q1-Q4 2025", f"{CHECK} 100%"),
        ("Branch Protection Configs", "Current", "All repos", f"{CHECK} 100%"),
        ("Secret Scan Results", "12 months", "12 months", f"{CHECK} 100%"),
        ("Access Review Records", "Quarterly", "Q1-Q4 2025", f"{WARNING} 96%"),
        ("Deprovisioning Logs", "All terminations", "All", f"{CHECK} 100%"),
        ("Contractor Documentation", "All contracts", "All", f"{CHECK} 100%"),
    ]
    
    for evidence_type, required, collected, completeness in evidence:
        ws[f'A{row}'] = evidence_type
        ws[f'B{row}'] = required
        ws[f'C{row}'] = collected
        ws[f'D{row}'] = completeness
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15


def create_approval_signoff_sheet(wb, styles):
    """Create Approval_Sign_Off sheet."""
    ws = wb["Approval_Sign_Off"]
    
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "DASHBOARD REVIEW & APPROVAL"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 40
    
    row = 3
    ws.merge_cells(f'A{row}:D{row+3}')
    cell = ws[f'A{row}']
    cert_text = f"""Source Code Security Compliance Dashboard Certification

Dashboard Period: {datetime.now().strftime('%B %Y')}
Overall Compliance Score: [See Executive_Summary]

I certify that this dashboard accurately represents the source code security posture
and that all data has been collected according to ISMS-POL-A.8.4 requirements."""
    cell.value = cert_text
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 100
    
    row += 5
    ws[f'A{row}'] = "CISO Approval:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Name: ________________"
    row += 1
    ws[f'A{row}'] = "Signature: ________________"
    row += 1
    ws[f'A{row}'] = "Date: ________________"
    
    ws.column_dimensions['A'].width = 40


# ============================================================================
# SECTION 8: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the Excel workbook."""
    logger.info("Starting ISMS-A84-3 Source Code Security Dashboard Generation...")
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("Creating Executive_Summary sheet...")
    create_executive_summary_sheet(wb, styles)
    
    logger.info("Creating Repository_Overview sheet...")
    create_repository_overview_sheet(wb, styles)
    
    logger.info("Creating Access_Control_Metrics sheet...")
    create_access_control_metrics_sheet(wb, styles)
    
    logger.info("Creating Branch_Protection_Metrics sheet...")
    create_branch_protection_metrics_sheet(wb, styles)
    
    logger.info("Creating Secret_Management_Metrics sheet...")
    create_secret_management_metrics_sheet(wb, styles)
    
    logger.info("Creating Third_Party_Access sheet...")
    create_third_party_access_sheet(wb, styles)
    
    logger.info("Creating Trend_Analysis sheet...")
    create_trend_analysis_sheet(wb, styles)
    
    logger.info("Creating Gap_Priority_Matrix sheet...")
    create_gap_priority_matrix_sheet(wb, styles)
    
    logger.info("Creating Action_Items sheet...")
    create_action_items_sheet(wb, styles)
    
    logger.info("Creating Evidence_Summary sheet...")
    create_evidence_summary_sheet(wb, styles)
    
    logger.info("Creating Approval_Sign_Off sheet...")
    create_approval_signoff_sheet(wb, styles)
    
    output_filename = f"ISMS-IMP-A.8.4.3_Source_Code_Security_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(output_filename)
    logger.info(f"\n✅ Dashboard generated successfully: {output_filename}")
    logger.info(f"{CHART} Sheets created: {len(wb.sheetnames)}")
    logger.info("\n🎯 Dashboard ready for executive review and compliance reporting!")


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
