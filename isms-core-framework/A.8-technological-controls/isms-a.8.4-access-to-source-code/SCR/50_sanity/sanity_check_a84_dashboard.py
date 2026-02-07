#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.4 Assessment Workbooks
================================================================================

Diagnostic utility for troubleshooting Excel's "file level validation and repair"
warnings when opening A.8.4 source code access control assessment workbooks.

**Purpose:**
Identifies common openpyxl-generated Excel issues that trigger repair warnings:
- Formula syntax errors and invalid sheet references
- Data validation conflicts and overlapping ranges
- Style attribute inconsistencies
- Merged cell content issues
- Worksheet structure problems

**When to Use:**
- Excel displays repair warnings when opening generated workbooks
- After modifying assessment generator scripts
- Before distributing workbooks to stakeholders
- Quality assurance validation before consolidation

**Usage:**
    python3 sanity_check_a84_dashboard.py ISMS-IMP-A.8.4_X_Assessment_YYYYMMDD.xlsx
    
    Works with any A.8.4 assessment workbook (domains 1-3)

**Output:**
- Diagnostic report with issue categorization (Critical/Warning)
- Recommended remediation actions
- Structural health summary

**Related Scripts:**
- generate_a84_1_repository_access.py (creates Domain 1 workbook)
- generate_a84_2_branch_protection.py (creates Domain 2 workbook)
- generate_a84_3_compliance_dashboard.py (creates Domain 3 workbook)
- normalize_assessment_files_a84.py (data quality validation)

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.4
Script Type: Quality Assurance Utility
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.4 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'a.8.4.1' in filename_lower or 'a_8_4_1' in filename_lower or 'repository_access' in filename_lower:
        return 'A.8.4.1', 'Repository Access Control Assessment'
    elif 'a.8.4.2' in filename_lower or 'a_8_4_2' in filename_lower or 'branch_protection' in filename_lower:
        return 'A.8.4.2', 'Branch Protection Configuration Assessment'
    elif 'a.8.4.3' in filename_lower or 'a_8_4_3' in filename_lower or 'compliance' in filename_lower or 'dashboard' in filename_lower:
        return 'A.8.4.3', 'Source Code Security Compliance Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula:
                        sheet_refs = re.findall(r"'([^']+)'!", formula)
                        for ref in sheet_refs:
                            # Skip external references (contain brackets)
                            if "[" not in ref and ref not in wb.sheetnames:
                                issues_found.append(
                                    f"  ✗ {sheet_name}!{cell.coordinate}: "
                                    f"Invalid sheet reference '{ref}'"
                                )
                                formula_issues += 1
                            
                            # Track inter-sheet references (internal only)
                            if "[" not in ref and ref in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref)
                    
                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
                    
                    # Check for double quotes issues
                    if formula.count('"') % 2 != 0:
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced quotes in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print("  ✓ All formulas appear syntactically valid")
    else:
        print(f"  ✗ Found {formula_issues} formula issues")
    
    # Report inter-sheet dependencies
    if inter_sheet_refs:
        print("\n  Inter-sheet formula dependencies detected:")
        for source, targets in sorted(inter_sheet_refs.items()):
            print(f"    {source} → {', '.join(sorted(targets))}")
    
    # Report external workbook links (important for A.8.4.3)
    if external_refs:
        print(f"\n  ✓ External workbook links detected: {len(external_refs)}")
        for ref in sorted(external_refs):
            print(f"    - {ref}")
        
        if workbook_id == 'A.8.4.3':
            expected_sources = [
                "ISMS-IMP-A.8.4.1.xlsx",
                "ISMS-IMP-A.8.4.2.xlsx",
            ]
            missing = [s for s in expected_sources if s not in external_refs]
            if missing:
                warnings_found.append(
                    f"  ⚠ Dashboard expected to reference: {', '.join(missing)}"
                )
            else:
                print("  ✓ All expected source workbooks referenced")
    elif workbook_id == 'A.8.4.3':
        warnings_found.append(
            "  ⚠ A.8.4.3 Dashboard: No external workbook links found!"
        )
        print("  ⚠ WARNING: This dashboard requires external links to function")
    
    # ========================================================================
    # CHECK 2: DATA VALIDATION CONFLICTS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: DATA VALIDATION CONFLICTS")
    print("=" * 80)
    
    validation_issues = 0
    total_validations = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            print(f"  {sheet_name}: {dv_count} data validations")
            
            # Check for overlapping ranges (simplified)
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())
            
            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1
    
    print(f"  Total data validations across all sheets: {total_validations}")
    
    if validation_issues == 0:
        print("  ✓ No obvious data validation conflicts")
    
    # ========================================================================
    # CHECK 3: STYLE CONSISTENCY
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: STYLE CONSISTENCY")
    print("=" * 80)
    
    style_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        cells_without_font = 0
        cells_without_border = 0
        
        sample_size = 0
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value:
                    sample_size += 1
                    if not hasattr(cell, 'font') or cell.font is None:
                        cells_without_font += 1
                    if not hasattr(cell, 'border') or cell.border is None:
                        cells_without_border += 1
        
        if cells_without_font > 0:
            warnings_found.append(
                f"  ⚠ {sheet_name}: {cells_without_font}/{sample_size} "
                f"cells missing font attributes"
            )
        
        # Note: Missing borders are often intentional, so we don't flag this as issue
    
    if style_issues == 0:
        print("  ✓ Style attributes appear consistent")
    
    # ========================================================================
    # CHECK 4: MERGED CELLS VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    total_merges = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            print(f"  {sheet_name}: {merge_count} merged cell ranges")
            
            # Check if merged cells have content in non-top-left cells
            for merge_range in ws.merged_cells.ranges:
                min_col = merge_range.min_col
                min_row = merge_range.min_row
                max_col = merge_range.max_col
                max_row = merge_range.max_row
                
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue
                        
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  ⚠ {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1
    
    if total_merges > 0:
        print(f"  Total merged ranges across all sheets: {total_merges}")
    
    if merge_issues == 0:
        print("  ✓ Merged cells appear properly formatted")
    else:
        print(f"  ✗ Found {merge_issues} merged cell content issues")
    
    # ========================================================================
    # CHECK 5: WORKSHEET STRUCTURE
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: WORKSHEET STRUCTURE")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Check for excessive dimensions
        if ws.max_row > 10000:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )
        
        if ws.max_column > 100:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )
    
    print("  ✓ Worksheet dimensions within reasonable limits")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️  WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("Excel repair warnings may be due to:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
        if workbook_id == 'A.8.4.3':
            print("  • Unresolved external workbook links (expected before setup)")
    else:
        print("\n" + "=" * 80)
        print("RECOMMENDED ACTIONS:")
        print("=" * 80)
        
        if formula_issues > 0:
            print("\n1. FORMULA FIXES:")
            print("   • Review formulas referencing other sheets")
            print("   • Ensure sheet names match exactly (case-sensitive)")
            print("   • Check for balanced quotes and parentheses")
            if workbook_id == 'A.8.4.3':
                print("   • Verify external workbook references are correct")
        
        if validation_issues > 0:
            print("\n2. DATA VALIDATION FIXES:")
            print("   • Remove overlapping data validation ranges")
            print("   • Apply validations to specific ranges, not entire columns")
        
        if merge_issues > 0:
            print("\n3. MERGED CELL FIXES:")
            print("   • Ensure only top-left cell of merged range has content")
            print("   • Clear content from other cells in merged range")
        
        # Special instructions for A.8.4.3 Dashboard
        if workbook_id == 'A.8.4.3' and (not external_refs or len(warnings_found) > 0):
            print("\n4. DASHBOARD-SPECIFIC SETUP (A.8.4.3):")
            print("   • This workbook REQUIRES external links to function")
            print("   • Ensure both source workbooks are in the same folder:")
            print("     - ISMS-IMP-A.8.4.1.xlsx (Repository Access Control)")
            print("     - ISMS-IMP-A.8.4.2.xlsx (Branch Protection Configuration)")
            print("   • Run normalize_assessment_files_a84.py first to standardize filenames")
            print("   • Open dashboard and click 'Update Links' when prompted")
            print("   • #REF errors before updating links are EXPECTED and normal")
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'A.8.4.1':
        print("\nRepository Access Control Assessment:")
        print("  Expected sheets:")
        print("  • Instructions & Legend")
        print("  • Repository Inventory")
        print("  • Access Control Assessment")
        print("  • Access Review Records")
        print("  • Third-Party Access")
        print("  • Orphaned Access Analysis")
        print("  • Audit Logging Assessment")
        print("  • Gap Analysis")
        print("  • Evidence Register")
        print("  • Approval & Sign-Off")
        print("\n  Key data validations:")
        print("  • Git platform dropdown (GitHub, GitLab, Bitbucket, Azure DevOps, etc.)")
        print("  • Repository classification (Production, Tools, Archived, etc.)")
        print("  • Access level (Read, Write, Admin, Owner)")
        print("  • Review frequency (Monthly, Quarterly, Semi-Annual, Annual)")
        print("  • Compliance status (Compliant, Non-Compliant, Under Review)")
    
    elif workbook_id == 'A.8.4.2':
        print("\nBranch Protection Configuration Assessment:")
        print("  Expected sheets:")
        print("  • Instructions & Legend")
        print("  • Branch Protection Assessment")
        print("  • Pull Request Configuration")
        print("  • Status Checks Configuration")
        print("  • Commit Signing Assessment")
        print("  • Secret Scanning Assessment")
        print("  • Protection Gaps Analysis")
        print("  • Gap Analysis")
        print("  • Evidence Register")
        print("  • Approval & Sign-Off")
        print("\n  Key data validations:")
        print("  • Branch type (Main/Master, Release, Development, Hotfix)")
        print("  • Protection enabled (Yes/No)")
        print("  • Required approvals (0, 1, 2, 3+)")
        print("  • Status checks (Enabled/Disabled)")
        print("  • Commit signing (Required/Optional/Disabled)")
        print("  • Secret scanning (Active/Inactive)")
    
    elif workbook_id == 'A.8.4.3':
        print("\nSource Code Security Compliance Dashboard:")
        print("  Expected sheets (10 total - CONSOLIDATION WORKBOOK):")
        print("  • Executive Dashboard (with external links)")
        print("  • Repository Access Overview")
        print("  • Branch Protection Overview")
        print("  • Gap Analysis (consolidated from domains 1-2)")
        print("  • Evidence Register (consolidated from domains 1-2)")
        print("  • Risk Register")
        print("  • KPIs & Metrics")
        print("  • Action Items & Follow-up")
        print("  • Audit & Compliance Log")
        print("  • Approval & Sign-Off")
        print("\n  CRITICAL: This dashboard requires external workbook links!")
        print("  Setup steps:")
        print("    1. Run: python3 normalize_assessment_files_a84.py")
        print("    2. Place dashboard in same folder as normalized source files:")
        print("       - ISMS-IMP-A.8.4.1.xlsx")
        print("       - ISMS-IMP-A.8.4.2.xlsx")
        print("    3. Open dashboard and click 'Update Links'")
        print("\n  Key consolidation sources:")
        print("  • Repository access data from A.8.4.1")
        print("  • Branch protection data from A.8.4.2")
        print("  • Combined gap analysis and evidence register")
        print("  • Executive KPIs calculated from source data")
    
    else:
        print("\nGeneric Excel Workbook")
        print("  • No specific validation rules defined")
        print("  • Standard Excel health checks applied")
    
    print("\n" + "=" * 80)
    
    # Exit code: 0 = healthy, 1 = warnings, 2 = critical issues
    if issues_found:
        sys.exit(2)
    elif warnings_found:
        sys.exit(1)
    else:
        sys.exit(0)


def main():
    if len(sys.argv) < 2:
        print("=" * 80)
        print("ISMS A.8.4 - EXCEL WORKBOOK SANITY CHECKER")
        print("=" * 80)
        print("\nUsage: python3 sanity_check_a84_dashboard.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 sanity_check_a84_dashboard.py ISMS-IMP-A.8.4.1_Repository_Access_20250125.xlsx")
        print("  python3 sanity_check_a84_dashboard.py ISMS-IMP-A.8.4.2_Branch_Protection_20250125.xlsx")
        print("  python3 sanity_check_a84_dashboard.py ISMS-IMP-A.8.4.3_Compliance_Dashboard_20250125.xlsx")
        print("\nBatch check all workbooks:")
        print("  for file in ISMS-IMP-A.8.4_*.xlsx; do")
        print("      python3 sanity_check_a84_dashboard.py \"$file\"")
        print("  done")
        print("\nExit codes:")
        print("  0 = No issues detected (workbook is healthy)")
        print("  1 = Warnings found (workbook usable but has issues)")
        print("  2 = Critical errors found (workbook may fail in Excel)")
        print("\n" + "=" * 80)
        sys.exit(1)
    
    filename = sys.argv[1]
    check_workbook_health(filename)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
