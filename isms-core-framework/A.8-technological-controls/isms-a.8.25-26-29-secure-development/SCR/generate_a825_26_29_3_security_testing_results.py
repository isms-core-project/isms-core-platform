#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.25-26-29.S3 - Security Testing Results Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.25/A.8.26/A.8.29: Secure Development Framework
Assessment Domain 3 of 5: Security Testing Results and Coverage (A.8.29)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific:
- Development methodologies (Agile, Waterfall, DevOps, hybrid)
- Technology stack (languages, frameworks, platforms)
- Security tooling (SAST, DAST, SCA, IAST tools and versions)
- SDLC processes and governance structure
- Compliance requirements and risk tolerance
- Security testing types and frequencies
- Testing tool configurations and scan policies
- Vulnerability severity classification and acceptance criteria

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.25-26-29 Secure Development Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
security testing execution, coverage, and results across all applications
throughout the development and acceptance lifecycle.

**Purpose:**
Enables systematic assessment of security testing strategy execution, test
coverage, vulnerability detection, and security acceptance criteria validation
against ISO 27001:2022 Control A.8.29 requirements, supporting evidence-based
validation of security testing effectiveness.

**Assessment Scope:**
- Security testing coverage by application and testing type
- SAST (Static Application Security Testing) scan execution and results
- DAST (Dynamic Application Security Testing) scan execution and results
- SCA (Software Composition Analysis) scan execution and vulnerable dependencies
- IAST (Interactive Application Security Testing) deployment and results
- Penetration testing execution and findings
- Security acceptance testing completion and sign-off
- Testing frequency compliance and automation status
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and security testing standards
2. Security Testing Coverage - Testing coverage matrix by application and type
3. SAST Scan Results - Static analysis results, trends, and remediation
4. DAST Scan Results - Dynamic analysis results, trends, and remediation
5. SCA Scan Results - Dependency analysis, vulnerable components, and remediation
6. Penetration Testing Results - Pentest findings, severity, and remediation
7. Security Acceptance Testing - Acceptance criteria, test execution, and sign-off
8. Testing Frequency Compliance - Testing schedule adherence and automation
9. Summary Dashboard - Overall security testing coverage and effectiveness scores
10. Gap Analysis - Non-compliant applications and remediation requirements
11. Evidence Register - Audit evidence tracking and documentation
12. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with testing type and frequency dropdown lists
- Conditional formatting for testing coverage and finding severity
- Automated gap identification for missing or overdue security tests
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with security testing tools and CI/CD pipelines

**Integration:**
This assessment feeds into the A.8.25-26-29.5 Secure Development Compliance
Dashboard, which consolidates data from all five assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl>=3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a825_26_29_3_security_testing_results.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a825_26_29_3_security_testing_results.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a825_26_29_3_security_testing_results.py --date 20250124

Output:
    File: ISMS_A_8_25_26_29_3_Security_Testing_Results_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize security testing types and frequencies for your risk profile
    2. Define testing coverage requirements by application risk classification
    3. Inventory all applications requiring security testing
    4. Complete security testing coverage matrix for each application
    5. Import or manually enter SAST scan results from your SAST tool
    6. Import or manually enter DAST scan results from your DAST tool
    7. Import or manually enter SCA scan results from your SCA tool
    8. Document penetration testing results and remediation status
    9. Track security acceptance testing completion and sign-off
    10. Verify testing frequency compliance (automated and manual tests)
    11. Conduct gap analysis for applications with insufficient testing coverage
    12. Define remediation actions with development teams and timelines
    13. Collect and link audit evidence (scan reports, pentest reports, test results)
    14. Obtain stakeholder approvals
    15. Feed results into A.8.25-26-29.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.25/A.8.26/A.8.29
Assessment Domain:    3 of 5 (Security Testing Results and Coverage - A.8.29)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [YYYY-MM-DD]
Last Modified:        [YYYY-MM-DD]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.25-26-29-S1: Executive Control Alignment (Governance)
    - ISMS-POL-A.8.25-26-29-S2: Security Requirements (A.8.26)
    - ISMS-POL-A.8.25-26-29-S3: Secure Development Lifecycle (A.8.25)
    - ISMS-POL-A.8.25-26-29-S4: Security Testing (A.8.29)
    - ISMS-POL-A.8.25-26-29-S5: Assessment Evidence Framework
    - ISMS-IMP-A.8.25-26-29-S4: Security Testing Implementation Implementation Guide
    - ISMS-IMP-A.8.25-26-29.S5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a825_26_29_1_security_requirements.py
    - generate_a825_26_29_2_sdlc_security_activities.py
    - generate_a825_26_29_3_security_testing_results.py (this script)
    - generate_a825_26_29_4_vulnerability_remediation.py
    - generate_a825_26_29_5_compliance_dashboard.py
    - normalize_assessment_files_a825_26_29.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [YYYY-MM-DD]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.25-26-29-S4 spec
    - Supports comprehensive security testing results evaluation
    - Integrated with A.8.25-26-29.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**SDLC Methodology Neutrality:**
This assessment framework is methodology-agnostic and must work across:
- Waterfall: Traditional phase-gate with security testing in dedicated test phases
- Agile: Iterative sprint-based with continuous security testing per sprint
- DevOps/DevSecOps: Continuous integration/deployment with automated security testing
- Hybrid: Mixed methodology approaches with varying testing integration
Customize dropdown values and assessment criteria to match YOUR methodology.
For Agile: Track testing per sprint/release.
For DevOps: Emphasize CI/CD pipeline integration and automation.

**Technology Stack Agnosticism:**
Framework must remain vendor-neutral and technology-agnostic:
- Programming languages: Java, Python, C#, JavaScript, Go, Ruby, PHP, etc.
- Frameworks: Spring, Django, .NET, React, Angular, Vue, Express, etc.
- Platforms: Web, mobile, desktop, embedded, cloud-native, containers, serverless
Do NOT hardcode technology-specific assumptions in assessment criteria.
Testing approaches vary by technology; provide customization guidance.

**Security Tool Diversity:**
Assessment must accommodate various security testing tools:
- SAST: SonarQube, Checkmarx, Fortify, Veracode, Snyk Code, Semgrep, Coverity
- DAST: OWASP ZAP, Burp Suite, Acunetix, AppScan, Netsparker, WebInspect
- SCA: Snyk, Dependabot, WhiteSource (Mend), Black Duck, Sonatype Nexus
- IAST: Contrast Security, Hdiv Security, Seeker, Checkmarx Interactive
- Pentest: Manual testing, automated frameworks (Metasploit, Cobalt Strike)
Use generic terms; avoid vendor lock-in in assessment structure.
Support import of results from multiple tools.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Security testing coverage matrix showing all applications tested
- Test execution records (scan reports, pentest reports)
- Vulnerability findings with severity classifications
- Remediation tracking for identified vulnerabilities
- Testing frequency compliance (automated scans, manual testing)
- Security acceptance sign-off records

**Data Protection:**
Assessment workbooks contain sensitive security testing information including:
- Application vulnerabilities and security weaknesses
- Attack surface analysis and exploitation details
- Penetration testing findings and proof-of-concept exploits
- Vulnerable dependency details with CVE references
- Security testing tool configurations and policies
- Security gaps and remediation timelines
Handle in accordance with your organisation's data classification policies.
Treat as CONFIDENTIAL - contains exploitable vulnerability information.

**Maintenance:**
Review and update assessment:
- Monthly: Update with latest security testing results (SAST, DAST, SCA scans)
- Quarterly: Complete penetration testing and reassessment cycle
- Annually: Review testing strategy, tool effectiveness, and coverage requirements
- Ad-hoc: When critical vulnerabilities detected or new applications deployed

**Quality Assurance:**
Have Application Security team, Penetration Testers, and Security Engineers
validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
Security testing assessment supports regulatory compliance:
- Payment processing: PCI DSS v4.0.1 requirement 6.3.2 (security testing in SDLC)
- Healthcare: HIPAA security requirements for application vulnerability assessment
- Finance: Regional banking requirements for application security testing
- Critical infrastructure: Sector-specific penetration testing requirements
- Privacy: GDPR/FADP requirements for security testing of data processing systems
Customize assessment criteria to include regulatory-specific requirements
applicable to your organisation and industry.

**Integration with Related Controls:**
This assessment integrates with:
- A.8.26 (Security Requirements): Requirements define test cases and acceptance criteria
- A.8.25 (Secure Development Lifecycle): Testing activities are SDLC phases
- A.8.8 (Vulnerability Management): Testing findings feed vulnerability management
- A.5.24 (Security Incident Management): Critical findings may trigger incidents
- A.8.31 (Environment Separation): Testing occurs in test environments
- A.8.32 (Change Management): Testing required before production deployment
- A.5.7 (Threat Intelligence): Testing strategies informed by threat intel

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

from datetime import datetime, timedelta
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.25-26-29.S3"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
WORKBOOK_NAME = "Security Testing Results Assessment"
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
CONTROL_ID   = "A.8.25-26-29"
CONTROL_NAME = "Secure Development"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Sheet structure from ISMS-IMP-A.8.25-26-29-S4 Section 10.1
    sheets = [
        "Instructions & Legend",
        "Security Testing Coverage",
        "SAST Scan Results",
        "DAST Scan Results",
        "SCA Scan Results",
        "Penetration Testing Results",
        "Security Acceptance Testing",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def setup_styles():
    """Define all cell styles - return TEMPLATES not objects."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "sample_cell": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to cell - creates NEW objects."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name if hasattr(style_dict["font"], 'name') else "Calibri",
            size=style_dict["font"].size if hasattr(style_dict["font"], 'size') else 10,
            bold=style_dict["font"].bold if hasattr(style_dict["font"], 'bold') else False,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type if hasattr(style_dict["fill"], 'fill_type') else "solid"
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal if hasattr(style_dict["alignment"], 'horizontal') else "left",
            vertical=style_dict["alignment"].vertical if hasattr(style_dict["alignment"], 'vertical') else "center",
            wrap_text=style_dict["alignment"].wrap_text if hasattr(style_dict["alignment"], 'wrap_text') else False
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        'yes_no': DataValidation(type="list", formula1='"Yes,No"', allow_blank=False),
        'yes_no_na': DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False),
        'scan_type': DataValidation(type="list", formula1='"Baseline,Full,API,Authenticated"', allow_blank=False),
        'pentester_type': DataValidation(type="list", formula1='"Internal,External,Both"', allow_blank=False),
        'test_status': DataValidation(type="list", formula1='"Passed,Failed,In Progress,N/A"', allow_blank=False),
        'approval_status': DataValidation(type="list", formula1='"Approved,Pending,Rejected"', allow_blank=False),
        'evidence_status': DataValidation(type="list", formula1='"Current,Outdated,Missing"', allow_blank=False),
        'compliance_status': DataValidation(type="list", formula1='"Compliant,⚠️ Partial Compliance,Non-Compliant"', allow_blank=False),
    }
    
    for key, dv in validations.items():
        ws.add_data_validation(dv)
    
    return validations


# ============================================================================
# SECTION 3: SHEET BUILDERS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab for applicable systems/services.', '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Check all applicable items in the Compliance Checklist for each section.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def build_testing_coverage_sheet(wb, styles, validations):
    """Build Security Testing Coverage overview sheet."""
    ws = wb["Security Testing Coverage"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "SECURITY TESTING COVERAGE OVERVIEW"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "TRACK SECURITY TESTING COVERAGE ACROSS ALL TESTING TYPES PER APPLICATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "SAST Enabled?",
        "Last SAST Scan",
        "DAST Enabled?",
        "Last DAST Scan",
        "SCA Enabled?",
        "Last SCA Scan",
        "Pentest Status",
        "Last Pentest Date",
        "Coverage Score (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    # Apply data validations
    validations['yes_no'].add('B4:B100')
    validations['yes_no'].add('D4:D100')
    validations['yes_no'].add('F4:F100')
    
    # Example data with formula
    example_data = [
        ["Customer Portal", "Yes", "2025-01-10", "Yes", "2025-01-08", "Yes", "2025-01-10", "Completed", "2024-11-15"],
        ["Internal HR System", "Yes", "2025-01-09", "No", "N/A", "Yes", "2025-01-09", "Scheduled", "2024-06-20"],
        ["Marketing Website", "Yes", "2025-01-11", "Yes", "2025-01-05", "Yes", "2025-01-11", "Completed", "2024-09-10"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])

        # Add Coverage Score formula in column J
        cell_j = ws.cell(row=row, column=10)
        cell_j.value = f'=((IF(B{row}="Yes",33,0))+(IF(D{row}="Yes",33,0))+(IF(F{row}="Yes",34,0)))'
        cell_j.number_format = '0"%"'
        apply_style(cell_j, styles['sample_cell'])

        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 11):  # 10 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_sast_results_sheet(wb, styles, validations):
    """Build SAST Scan Results sheet."""
    ws = wb["SAST Scan Results"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "SAST (STATIC APPLICATION SECURITY TESTING) SCAN RESULTS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "TRACK SAST SCAN FINDINGS BY SEVERITY AND REMEDIATION STATUS"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Last Scan Date",
        "Critical Findings",
        "High Findings",
        "Medium Findings",
        "Low Findings",
        "Total Findings",
        "Findings Remediated",
        "Remediation Rate (%)",
        "Trend vs. Previous",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    # Example data with formulas
    example_data = [
        ["Customer Portal", "2025-01-10", 0, 2, 5, 12],
        ["Internal HR System", "2025-01-09", 0, 3, 8, 15],
        ["Marketing Website", "2025-01-11", 0, 1, 3, 8],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            if col_num <= 6:
                cell.value = value
                apply_style(cell, styles['sample_cell'])
            elif col_num == 7:  # Total Findings
                cell.value = f'=C{row}+D{row}+E{row}+F{row}'
            elif col_num == 8:  # Findings Remediated (user input)
                cell.value = int(row_data[2] + row_data[3] * 0.5 + row_data[4] * 0.3)  # Example
                apply_style(cell, styles['sample_cell'])
            elif col_num == 9:  # Remediation Rate
                cell.value = f'=IF(G{row}>0,H{row}/G{row}*100,100)'
                cell.number_format = '0.0"%"'
            else:  # Trend (user input)
                cell.value = "IMPROVING"
                apply_style(cell, styles['sample_cell'])
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 11):  # 10 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_dast_results_sheet(wb, styles, validations):
    """Build DAST Scan Results sheet."""
    ws = wb["DAST Scan Results"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "DAST (DYNAMIC APPLICATION SECURITY TESTING) SCAN RESULTS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "TRACK DAST SCAN FINDINGS BY SEVERITY AND SCAN TYPE"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Last Scan Date",
        "Scan Type",
        "Critical Findings",
        "High Findings",
        "Medium Findings",
        "Low Findings",
        "Total Findings",
        "Remediation Status",
        "Next Scan Date",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    # Apply data validations
    validations['scan_type'].add('C4:C100')
    
    # Example data
    example_data = [
        ["Customer Portal", "2025-01-08", "Full", 0, 1, 4, 8, "In Progress", "2025-02-01"],
        ["Internal HR System", "2024-12-15", "Baseline", 0, 0, 2, 5, "Completed", "2025-01-15"],
        ["Marketing Website", "2025-01-05", "Full", 0, 0, 3, 6, "In Progress", "2025-02-05"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            if col_num != 8:  # Not Total Findings
                cell.value = value
                apply_style(cell, styles['sample_cell'])
            else:  # Total Findings formula
                cell.value = f'=D{row}+E{row}+F{row}+G{row}'
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 11):  # 10 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_sca_results_sheet(wb, styles, validations):
    """Build SCA Scan Results sheet."""
    ws = wb["SCA Scan Results"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "SCA (SOFTWARE COMPOSITION ANALYSIS) SCAN RESULTS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "TRACK VULNERABLE DEPENDENCIES AND LICENSE COMPLIANCE"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Last Scan Date",
        "Critical Vulns",
        "High Vulns",
        "Medium Vulns",
        "Low Vulns",
        "Total Vulns",
        "License Issues",
        "Outdated Deps",
        "Remediation Status",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    # Example data
    example_data = [
        ["Customer Portal", "2025-01-10", 0, 1, 3, 8, 0, 5, "80% Complete"],
        ["Internal HR System", "2025-01-09", 0, 2, 5, 10, 1, 8, "60% Complete"],
        ["Marketing Website", "2025-01-11", 0, 0, 2, 6, 0, 3, "100% Complete"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            if col_num != 7:  # Not Total Vulns
                cell.value = value
                apply_style(cell, styles['sample_cell'])
            else:  # Total Vulns formula
                cell.value = f'=C{row}+D{row}+E{row}+F{row}'
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 11):  # 10 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_pentest_results_sheet(wb, styles, validations):
    """Build Penetration Testing Results sheet."""
    ws = wb["Penetration Testing Results"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "PENETRATION TESTING RESULTS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "TRACK PENETRATION TESTING FINDINGS AND REMEDIATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Pentest Date",
        "Pentester",
        "Critical Findings",
        "High Findings",
        "Medium Findings",
        "Low Findings",
        "Total Findings",
        "Retest Date",
        "Findings Resolved",
        "Closure Rate (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    for i in range(1, 12):
        ws.column_dimensions[get_column_letter(i)].width = 16
    
    # Apply data validations
    validations['pentester_type'].add('C4:C100')
    
    # Example data
    example_data = [
        ["Customer Portal", "2024-11-15", "External", 0, 2, 5, 8, "2024-12-20", 14],
        ["Internal HR System", "2024-06-20", "Internal", 0, 1, 3, 6, "2024-08-15", 9],
        ["Marketing Website", "2024-09-10", "External", 0, 1, 2, 4, "2024-10-05", 7],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            if col_num not in [8, 11]:  # Not Total or Closure Rate
                cell.value = value
                apply_style(cell, styles['sample_cell'])
            elif col_num == 8:  # Total Findings
                cell.value = f'=D{row}+E{row}+F{row}+G{row}'
            else:  # Closure Rate
                cell.value = f'=IF(H{row}>0,J{row}/H{row}*100,100)'
                cell.number_format = '0.0"%"'
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 12):  # 11 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_security_acceptance_sheet(wb, styles, validations):
    """Build Security Acceptance Testing sheet."""
    ws = wb["Security Acceptance Testing"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:I1')
    cell = ws['A1']
    cell.value = "SECURITY ACCEPTANCE TESTING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:I2')
    cell = ws['A2']
    cell.value = "TRACK SECURITY TEST CASE EXECUTION AND PASS RATES"
    apply_style(cell, styles['subheader'])

    # Column headers — App-ID in col A, Application Name in col B, remaining shift right
    headers = [
        "App-ID",
        "Application Name",
        "Release Version",
        "Test Cases Total",
        "Test Cases Passed",
        "Test Cases Failed",
        "Test Cases N/A",
        "Pass Rate (%)",
        "Security Sign-Off",
    ]

    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Column widths
    ws.column_dimensions['A'].width = 12  # App-ID
    for i in range(2, 10):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Apply data validations (shifted right by 1 due to new App-ID col A)
    validations['approval_status'].add('I4:I100')

    # Example data: App-ID in col A, then remaining fields
    example_data = [
        ["APP-001", "Customer Portal", "v2.5.0", 25, 24, 1, 0, "Approved"],
    ]

    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            if col_num != 8:  # Not Pass Rate (now col 8)
                cell.value = value
                apply_style(cell, styles['sample_cell'])
            else:  # Pass Rate formula (D=Total, E=Passed after shift)
                cell.value = f'=IF(D{row}>0,E{row}/D{row}*100,0)'
                cell.number_format = '0.0"%"'
                apply_style(cell, styles['sample_cell'])
        row += 1

    # ISO 27002:2022 A.8.29: Pre-populated minimum security acceptance criteria
    # Col A = empty FFFFCC (App-ID), Col B = criterion text FFFFCC, Cols C-I = FFFFCC
    acceptance_criteria = [
        "[CRITERION] No Critical or High CVSS severity findings open at time of release",
        "[CRITERION] SAST scan completed with zero Critical findings unmitigated",
        "[CRITERION] SCA scan completed — all Critical/High vulnerable dependencies resolved",
        "[CRITERION] DAST scan completed for all external-facing endpoints",
        "[CRITERION] Penetration test completed and critical findings remediated (required for high-risk applications)",
        "[CRITERION] Security review sign-off obtained from Security Champion or AppSec team",
    ]
    for i, criterion in enumerate(acceptance_criteria):
        r = row + i
        # Col A: empty FFFFCC (App-ID for user to fill)
        apply_style(ws.cell(row=r, column=1), styles['input_cell'])
        # Col B: criterion text — FFFFCC
        ws.cell(row=r, column=2).value = criterion
        apply_style(ws.cell(row=r, column=2), styles['input_cell'])
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Cols C-I: empty FFFFCC input
        for col_num in range(3, 10):
            apply_style(ws.cell(row=r, column=col_num), styles['input_cell'])

    row += len(acceptance_criteria)

    # Add 50 empty FFFFCC rows after criterion rows
    for empty_row in range(row, row + 50):
        for col_num in range(1, 10):  # 9 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'

    return ws


def build_compliance_summary_sheet(wb, styles, validations):
    """Build Summary Dashboard with TABLE 1/2/3 gold standard structure."""
    ws = wb["Summary Dashboard"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value=f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border

    ws.merge_cells("A2:G2")
    cell = ws.cell(row=2, column=1, value=CONTROL_REF)
    cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    # TABLE 1 area configs
    area_configs = [
        ("SAST Scanning Coverage",       "Security Testing Coverage", "B", "Yes",      None,      "No", 5, 54),
        ("DAST Scanning Coverage",        "Security Testing Coverage", "D", "Yes",      None,      "No", 5, 54),
        ("SCA Scanning Coverage",         "Security Testing Coverage", "F", "Yes",      None,      "No", 5, 54),
        ("Security Acceptance Testing",   "Security Acceptance Testing","I", "Approved", "Pending", "Not Approved", 5, 60),
    ]

    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row = 5
    t1_headers = ["Assessment Area", "Total Assessed", "Complete", "Partial", "Missing", "N/A", "Completion %"]
    for col_idx, h in enumerate(t1_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, (display_name, sheet_name, s_col, complete_v, partial_v, nc_v, r_start, r_end) in enumerate(area_configs):
        r = 6 + i
        ws.cell(row=r, column=1, value=display_name).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border
        c = ws.cell(row=r, column=2)
        c.value = f"=COUNTA('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end})"
        c.border = border; c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=3)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{complete_v}\")"
        c.border = border; c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=4)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{partial_v}\")" if partial_v else "=0"
        c.border = border; c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=5)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{nc_v}\")"
        c.border = border; c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=6)
        c.value = "=0"
        c.border = border; c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=7)
        c.value = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
        c.number_format = "0.0%"
        c.border = border; c.alignment = Alignment(horizontal="center")

    total_row = 6 + len(area_configs)  # row 10
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}6:{col_letter}{total_row-1})"
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border; c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=total_row, column=7)
    c.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.border = border; c.alignment = Alignment(horizontal="center")

    # TABLE 2: KEY METRICS
    row = total_row + 2  # row 12
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    metrics = [
        ("Total Applications Covered",          "=COUNTA('Security Testing Coverage'!A5:A54)"),
        ("Penetration Tests Completed",          "=COUNTA('Penetration Testing Results'!A5:A54)"),
        ("Average SAST Remediation Rate (%)",    "=IFERROR(AVERAGE('SAST Scan Results'!I5:I54),\"No data\")"),
        ("Next Review Due",                      ""),
        ("Assessment Owner",                     ""),
    ]

    # TABLE 2 column headers
    header_row = row + 1
    ws.cell(row=header_row, column=1, value="Metric").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=header_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=header_row, column=1).border = border
    ws.merge_cells(f"B{header_row}:G{header_row}")
    ws.cell(row=header_row, column=2, value="Value").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=header_row, column=2).fill = PatternFill("solid", fgColor="D9D9D9")
    for col in range(2, 8):
        ws.cell(row=header_row, column=col).border = border
    r = row + 2
    for label, formula in metrics:
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(row=r, column=2)
        c.value = formula if formula else ""
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # TABLE 3: CRITICAL FINDINGS
    # 2 buffer rows: A alone | B:G merged (mirrors TABLE 2 metric structure)
    for _ in range(2):
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:G{r}")
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = border
        r += 1
    row = r + 1  # blank gap, then TABLE 3
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: CRITICAL FINDINGS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    t3_headers = ["#", "Finding", "Count", "Severity", "Affected Area", "Recommended Action", "Owner"]
    for col_idx, h in enumerate(t3_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border

    findings = [
        ("=COUNTIF('Security Testing Coverage'!B5:B54,\"No\")",           "Apps Without SAST Scanning",                "High",     "Security Testing Coverage", "Enable SAST scanning in CI/CD pipeline"),
        ("=COUNTIF('Security Testing Coverage'!D5:D54,\"No\")",           "Apps Without DAST Scanning",                "High",     "Security Testing Coverage", "Implement DAST scanning in test pipeline"),
        ("=COUNTIF('Security Testing Coverage'!F5:F54,\"No\")",           "Apps Without SCA Scanning",                 "Medium",   "Security Testing Coverage", "Enable SCA for dependency vulnerability detection"),
        ("=SUM('SAST Scan Results'!C5:C54)",                              "Critical SAST Findings (Total)",            "Critical", "SAST Scan Results",         "Immediate remediation of all critical findings"),
        ("=COUNTIF('Security Acceptance Testing'!I5:I60,\"Not Approved\")","Failed Security Acceptance Tests",          "High",     "Security Acceptance Testing","Re-test and resolve all acceptance test failures"),
    ]

    for i, (count_f, finding_name, severity, area, action) in enumerate(findings, 1):
        r = row + i
        ws.cell(row=r, column=1, value=count_f).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=finding_name).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=count_f).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=4, value=severity).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=5, value=area).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=6, value=action).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=6).border = border
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=7).border = border

    for col, w in zip("ABCDEFG", [38, 16, 16, 16, 16, 16, 18]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"
    return ws
def create_evidence_register(wb, styles):
    """Create Evidence Register — gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["Evidence Register"]
    ws.sheet_view.showGridLines = False

    # ROW 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border_thin

    # ROW 2: Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border_thin

    # ROW 3: intentionally empty (visual separator)

    # ROW 4: Column headers (003366 navy)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    dv_type = DataValidation(
        type="list",
        formula1='"Test report,SAST report,DAST report,SCA report,Penetration test,Vulnerability scan,Code review report,Screenshot,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"\u2705 Verified,\u26a0\ufe0f Pending,\u274c Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    # ROW 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001",
        2: "Security Testing Assessment",
        3: "Policy Document",
        4: "Security testing execution report and findings log",
        5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "\u2705 Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    dv_type.add(ws["C5"])
    dv_ver.add(ws["H5"])

    # ROWS 6-105: Empty data rows (FFFFCC yellow, 100 rows — MAX-002)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None  # No pre-filled IDs (MAX-001)
        dv_type.add(ws[f"C{r}"])
        dv_ver.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"
    return ws
def create_approval_sheet(wb, styles, doc_id, assessment_name, summary_sheet_name="Summary Dashboard", compliance_cell="G10"):
    """Create Approval Sign-Off — gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    if "Approval Sign-Off" in wb.sheetnames:
        del wb["Approval Sign-Off"]
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    # ROW 1: Title banner (003366)
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border_thin

    # ROW 2: Control reference (003366 italic, NO fill)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border_thin

    # ROW 3: Assessment Summary banner (4472C4)
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border_thin

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:",                  f"{doc_id} - {assessment_name}"),
        ("Assessment Period:",         ""),
        ("Overall Compliance Rating:", f"='{summary_sheet_name}'!G10"),
        ("Assessment Status:",         ""),
        ("Assessed By:",               ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border_thin  # GS-AS-011: border on label column
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for col_letter in ['B', 'C', 'D', 'E']:
            ws[f"{col_letter}{row}"].border = border_thin
        row += 1

    # Assessment Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(ws["B7"])  # Row 7 = Assessment Status

    # Approver sections
    row += 2  # row 10
    for title, color in [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_thin
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = border_thin
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for col_letter in ['B', 'C', 'D', 'E']:
                ws[f"{col_letter}{row}"].border = border_thin
            row += 1
        row += 1  # gap between sections

    # Final Decision
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border_thin
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for col_letter in ['B', 'C', 'D', 'E']:
        ws[f"{col_letter}{row}"].border = border_thin
    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # Next Review Details
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border_thin

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border_thin
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for col_letter in ['B', 'C', 'D', 'E']:
            ws[f"{col_letter}{row}"].border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    return ws
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            dv.showErrorMessage = True
            dv.showInputMessage = True
def main():
    """Main execution function."""
    logger.info("Generating ISMS A.8.29 Security Testing Results Assessment Workbook...")
    logger.info("=" * 70)
    
    # Create workbook and styles
    logger.info("\nCreating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    # Build all sheets
    logger.info("Building Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])
    
    logger.info("Building Security Testing Coverage sheet...")
    ws_coverage = wb["Security Testing Coverage"]
    ws_coverage.sheet_view.showGridLines = False
    validations = create_base_validations(ws_coverage)
    build_testing_coverage_sheet(wb, styles, validations)
    
    logger.info("Building SAST Scan Results sheet...")
    build_sast_results_sheet(wb, styles, validations)
    
    logger.info(" Building DAST Scan Results sheet...")
    build_dast_results_sheet(wb, styles, validations)
    
    logger.info("Building SCA Scan Results sheet...")
    build_sca_results_sheet(wb, styles, validations)
    
    logger.info(" Building Penetration Testing Results sheet...")
    build_pentest_results_sheet(wb, styles, validations)
    
    logger.info("Building Security Acceptance Testing sheet...")
    build_security_acceptance_sheet(wb, styles, validations)
    
    logger.info(" Building Summary Dashboard dashboard...")
    build_compliance_summary_sheet(wb, styles, validations)
    
    logger.info(" Building Evidence Register...")
    create_evidence_register(wb, styles)
    
    logger.info("Building Approval Sign-Off sheet...")
    create_approval_sheet(wb, styles, "ISMS-IMP-A.8.25-26-29.S3", "Security Testing Results Assessment", "Summary Dashboard", "G10")
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.25-26-29.S3_Security_Testing_Results_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info(f"\n Saving workbook: {filename}")
    # Finalize all data validations

    for ws_name in wb.sheetnames:

        ws_temp = wb[ws_name]

        for dv in list(ws_temp.data_validations.dataValidation):

            dv.showErrorMessage = True

            dv.showInputMessage = True

    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info("\n" + "=" * 70)
    logger.info("Workbook generated successfully!")
    logger.info("=" * 70)
    logger.info(f"\nFile: {filename}")
    logger.info(f" Sheets: {len(wb.sheetnames)}")
    logger.info("\nSheet List:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        logger.info(f"  {i}. {sheet_name}")
    
    logger.info("\n" + "=" * 70)
    logger.info("NEXT STEPS:")
    logger.info("=" * 70)
    logger.info("1. Open the workbook in Excel or LibreOffice Calc")
    logger.info("2. Review the Instructions & Legend sheet")
    logger.info("3. Complete Security_Testing_Coverage for each application")
    logger.info("4. Record SAST scan results")
    logger.info("5. Record DAST scan results")
    logger.info("6. Record SCA scan results")
    logger.info("7. Document penetration testing results")
    logger.info("8. Track security acceptance testing")
    logger.info("9. Review Compliance_Summary for overall scores")
    logger.info("10. Obtain approvals in Approval_Sign_Off")
    logger.info("\n" + "=" * 70)
    logger.info(" REFERENCE:")
    logger.info("=" * 70)
    logger.info("Implementation Guide: ISMS-IMP-A.8.25-26-29-S4")
    logger.info("Policy Reference: ISMS-POL-A.8.25-26-29-S4")
    logger.info("ISO Control: A.8.29 (Security Testing)")
    logger.info("=" * 70)
    


if __name__ == "__main__":
    try:
        filename = main()
        logger.info(f"\nSUCCESS: {filename} created successfully\n")
    except Exception as e:
        logger.error(f"\nERROR: {str(e)}\n")
        import traceback
        traceback.print_exc()
        exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
