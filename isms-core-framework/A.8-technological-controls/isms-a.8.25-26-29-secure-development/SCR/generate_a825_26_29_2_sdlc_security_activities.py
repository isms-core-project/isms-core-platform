#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.25-26-29.S2 - SDLC Security Activities Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.25/A.8.26/A.8.29: Secure Development Framework
Assessment Domain 2 of 5: Secure Development Lifecycle Activities (A.8.25)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific:
- Development methodologies (Agile, Waterfall, DevOps, hybrid)
- Technology stack (languages, frameworks, platforms)
- Security tooling (SAST, DAST, SCA, IAST tools and versions)
- SDLC processes and governance structure
- Compliance requirements and risk tolerance
- SDLC phase definitions and security gates
- Secure coding standards and code review processes
- Security training programs and certification requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.25-26-29 Secure Development Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
security activities integrated throughout the Software Development Lifecycle
(SDLC) across all development projects and teams.

**Purpose:**
Enables systematic assessment of security integration in SDLC phases, secure
coding practices, code review execution, security training, and security tool
deployment against ISO 27001:2022 Control A.8.25 requirements, supporting
evidence-based validation of secure development practices.

**Assessment Scope:**
- Security activities by SDLC phase (requirements, design, development, testing, deployment)
- Secure coding standard adoption and compliance
- Code review execution and security-focused review coverage
- Security training completion rates for developers
- Security tool deployment and integration (SAST, SCA, secret scanning, IDE plugins)
- Security champion program effectiveness
- Security defect management and remediation
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and SDLC security standards
2. SDLC Phase Activities - Security activities execution by phase and project
3. Secure Coding Compliance - Adoption of secure coding standards by language/framework
4. Code Review Tracking - Code review execution and security coverage
5. Security Training Status - Developer training completion and certification
6. Security Tool Deployment - SAST/SCA/secret scanning tool deployment status
7. Security Champion Program - Security champion coverage and effectiveness
8. Summary Dashboard - Overall SDLC security activity completion scores
9. Gap Analysis - Non-compliant projects/teams and remediation requirements
10. Evidence Register - Audit evidence tracking and documentation
11. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with SDLC phase and activity dropdown lists
- Conditional formatting for activity completion status
- Automated gap identification for missing security activities
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with code repositories, CI/CD pipelines, and training systems

**Integration:**
This assessment feeds into the A.8.25-26-29.5 Secure Development Compliance
Dashboard, which consolidates data from all five assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl>=3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a825_26_29_2_sdlc_security_activities.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a825_26_29_2_sdlc_security_activities.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a825_26_29_2_sdlc_security_activities.py --date 20250124

Output:
    File: ISMS_A_8_25_26_29_2_SDLC_Security_Activities_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize SDLC phase definitions to match your methodology
    2. Define security activities required for each SDLC phase
    3. Inventory all active development projects and teams
    4. Complete SDLC phase activity tracking for each project
    5. Assess secure coding standard adoption by development team
    6. Track code review execution and security-focused review coverage
    7. Verify security training completion for all developers
    8. Document security tool deployment (SAST, SCA, secret scanning)
    9. Conduct gap analysis for projects missing security activities
    10. Define remediation actions with development leads and timelines
    11. Collect and link audit evidence (code review records, training certificates, tool configs)
    12. Obtain stakeholder approvals
    13. Feed results into A.8.25-26-29.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.25/A.8.26/A.8.29
Assessment Domain:    2 of 5 (Secure Development Lifecycle Activities - A.8.25)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [YYYY-MM-DD]
Last Modified:        [YYYY-MM-DD]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.25-26-29-S1: Executive Control Alignment (Governance)
    - ISMS-POL-A.8.25-26-29-S2: Security Requirements (A.8.26)
    - ISMS-POL-A.8.25-26-29-S3: Secure Development Lifecycle (A.8.25)
    - ISMS-POL-A.8.25-26-29-S4: Security Testing (A.8.29)
    - ISMS-POL-A.8.25-26-29-S5: Assessment Evidence Framework
    - ISMS-IMP-A.8.25-26-29-S2: SDLC Security Integration Implementation Guide
    - ISMS-IMP-A.8.25-26-29.S5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a825_26_29_1_security_requirements.py
    - generate_a825_26_29_2_sdlc_security_activities.py (this script)
    - generate_a825_26_29_3_security_testing_results.py
    - generate_a825_26_29_4_vulnerability_remediation.py
    - generate_a825_26_29_5_compliance_dashboard.py
    - normalize_assessment_files_a825_26_29.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [YYYY-MM-DD]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.25-26-29-S2 spec
    - Supports comprehensive SDLC security activities evaluation
    - Integrated with A.8.25-26-29.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**SDLC Methodology Neutrality:**
This assessment framework is methodology-agnostic and must work across:
- Waterfall: Traditional phase-gate development with formal security reviews
- Agile: Iterative sprint-based development with security in each sprint
- DevOps/DevSecOps: Continuous integration/deployment with automated security
- Hybrid: Mixed methodology approaches with varying security integration
Customize dropdown values and assessment criteria to match YOUR methodology.
For Agile: Map activities to sprints, not traditional phases.
For DevOps: Emphasize automation and CI/CD pipeline security.

**Technology Stack Agnosticism:**
Framework must remain vendor-neutral and technology-agnostic:
- Programming languages: Java, Python, C#, JavaScript, Go, Ruby, PHP, etc.
- Frameworks: Spring, Django, .NET, React, Angular, Vue, Express, etc.
- Platforms: Web, mobile, desktop, embedded, cloud-native, serverless, etc.
Do NOT hardcode technology-specific assumptions in assessment criteria.
Secure coding standards vary by language; provide customization guidance.

**Security Tool Diversity:**
Assessment must accommodate various security tools in the SDLC:
- SAST: SonarQube, Checkmarx, Fortify, Veracode, Snyk Code, Semgrep
- SCA: Snyk, Dependabot, WhiteSource, Black Duck, Mend
- Secret Scanning: GitGuardian, TruffleHog, GitHub Secret Scanning
- IDE Plugins: SonarLint, Snyk IDE plugins, security-focused linters
- Code Review: GitHub, GitLab, Bitbucket, Gerrit, Crucible
Use generic terms; avoid vendor lock-in in assessment structure.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Documented security activities for each SDLC phase
- Evidence of secure coding standard adoption (policies, training)
- Code review records showing security-focused reviews
- Training completion records for all developers
- Security tool deployment and integration evidence
- Security defect tracking and remediation records

**Data Protection:**
Assessment workbooks contain sensitive development information including:
- Project names and development team structures
- Code review findings and security issues
- Training records and developer competency data
- Security tool configurations and findings
- Security gaps and remediation plans
Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Complete reassessment of all active projects
- Semi-annually: Update secure coding standards based on new threats
- Annually: Review SDLC security activities and adjust as needed
- Ad-hoc: When new projects start or methodologies change

**Quality Assurance:**
Have Application Security team, Development Leads, and DevOps Engineers
validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
SDLC security activities assessment supports regulatory compliance:
- Payment processing: PCI DSS v4.0.1 requirement 6.3 (secure SDLC practices)
- Healthcare: HIPAA security requirements for software development
- Finance: Regional banking requirements for secure development
- Critical infrastructure: Sector-specific secure development requirements
Customize assessment criteria to include regulatory-specific requirements
applicable to your organisation and industry.

**Integration with Related Controls:**
This assessment integrates with:
- A.8.4 (Access to Source Code): SDLC activities include access controls
- A.8.26 (Security Requirements): Requirements inform SDLC activities
- A.8.28 (Secure Coding): Secure coding standards are SDLC activities
- A.8.29 (Security Testing): Testing activities are part of SDLC phases
- A.8.31 (Environment Separation): SDLC phases map to environments
- A.8.32 (Change Management): SDLC activities feed change process
- A.5.24 (Security Incident Management): Defect management integrates

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

from datetime import datetime, timedelta
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.25-26-29.S2"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
WORKBOOK_NAME = "SDLC Security Activities Assessment"
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
CONTROL_ID   = "A.8.25-26-29"
CONTROL_NAME = "Secure Development"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Sheet structure from ISMS-IMP-A.8.25-26-29-S2 Section 7.2
    sheets = [
        "Instructions & Legend",
        "SDLC Phase Activities",
        "Secure Coding Standards",
        "Code Review Metrics",
        "Security Tools Deployment",
        "Security Tools Usage",
        "Developer Training",
        "Security Defect Mgmt",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def setup_styles():
    """Define all cell styles - return TEMPLATES not objects."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "sample_cell": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True),
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True),
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to cell - creates NEW objects."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name if hasattr(style_dict["font"], 'name') else "Calibri",
            size=style_dict["font"].size if hasattr(style_dict["font"], 'size') else 10,
            bold=style_dict["font"].bold if hasattr(style_dict["font"], 'bold') else False,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type if hasattr(style_dict["fill"], 'fill_type') else "solid"
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal if hasattr(style_dict["alignment"], 'horizontal') else "left",
            vertical=style_dict["alignment"].vertical if hasattr(style_dict["alignment"], 'vertical') else "center",
            wrap_text=style_dict["alignment"].wrap_text if hasattr(style_dict["alignment"], 'wrap_text') else False
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        'yes_no': DataValidation(type="list", formula1='"Yes,No"', allow_blank=False),
        'yes_no_na': DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False),
        'activity_status': DataValidation(type="list", formula1='"Complete,In Progress,Not Done,N/A"', allow_blank=False),
        'sdlc_methodology': DataValidation(type="list", formula1='"Waterfall,Agile,Scrum,DevOps,DevSecOps,Hybrid"', allow_blank=False),
        'tool_status': DataValidation(type="list", formula1='"Deployed,In Progress,Not Deployed,N/A"', allow_blank=False),
        'evidence_status': DataValidation(type="list", formula1='"Current,Outdated,Missing"', allow_blank=False),
        'compliance_status': DataValidation(type="list", formula1='"Compliant,⚠️ Partial Compliance,Non-Compliant"', allow_blank=False),
    }
    
    for key, dv in validations.items():
        ws.add_data_validation(dv)
    
    return validations


# ============================================================================
# SECTION 3: SHEET BUILDERS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab for applicable systems/services.', '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Check all applicable items in the Compliance Checklist for each section.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def build_sdlc_phase_activities_sheet(wb, styles, validations):
    """Build SDLC Phase Activities matrix sheet."""
    ws = wb["SDLC Phase Activities"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = "SDLC SECURITY ACTIVITIES BY PHASE"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:N2')
    cell = ws['A2']
    cell.value = "TRACK SECURITY ACTIVITY COMPLETION FOR EACH SDLC PHASE PER APPLICATION"
    apply_style(cell, styles['subheader'])

    # Column headers — App-ID in col A, Application Name in col B, remaining shift right
    headers = [
        "App-ID",
        "Application Name",
        "SDLC Methodology",
        "Requirements: Sec Req Defined",
        "Requirements: Risk Classification",
        "Design: Threat Modeling",
        "Design: Architecture Review",
        "Development: SAST Enabled",
        "Development: Code Review",
        "Testing: DAST Scan",
        "Testing: Security Testing",
        "Deployment: Security Checklist",
        "Maintenance: Vulnerability Monitoring",
        "Phase Compliance (%)",
    ]

    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Column widths
    ws.column_dimensions['A'].width = 12  # App-ID
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Apply data validations (shifted right by 1 due to new App-ID col A)
    validations['sdlc_methodology'].add('C4:C100')
    validations['activity_status'].add('D4:M100')

    # Example data: App-ID in col A, then remaining fields
    example_data = [
        ["APP-001", "Customer Portal", "Agile", "Complete", "Complete", "Complete", "Complete", "Complete", "Complete", "Complete", "Complete", "Complete", "Complete"],
    ]

    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])

        # Add Phase Compliance formula in column N (shifted from M)
        cell_n = ws.cell(row=row, column=14)
        cell_n.value = f'=(COUNTIF(D{row}:M{row},"Complete")/COUNTIF(D{row}:M{row},"<>N/A"))*100'
        cell_n.number_format = '0"%"'
        apply_style(cell_n, styles['sample_cell'])

        row += 1

    # ISO 27002:2022 A.8.25: Pre-populated required security activities per SDLC phase
    # Col A = empty FFFFCC (App-ID), Col B = activity description FFFFCC, Cols C-N = FFFFCC
    sdlc_activities = [
        "[REQUIREMENTS] Security requirements documented alongside functional requirements",
        "[REQUIREMENTS] Threat modelling initiated for high-risk applications",
        "[REQUIREMENTS] Data classification and privacy requirements identified",
        "[REQUIREMENTS] Regulatory and compliance requirements mapped to security controls",
        "[REQUIREMENTS] Security acceptance criteria defined",
        "[DESIGN] Security architecture review conducted before development begins",
        "[DESIGN] Secure design patterns applied (authentication, authorisation, encryption, logging)",
        "[DESIGN] Attack surface analysis performed",
        "[DESIGN] Security design reviewed by Security Champion or AppSec team",
        "[IMPLEMENTATION] Secure coding standards applied (OWASP Top 10, CWE Top 25)",
        "[IMPLEMENTATION] SAST scan integrated and run on every commit/PR",
        "[IMPLEMENTATION] SCA / dependency scanning enabled — no Critical/High unresolved vulnerabilities",
        "[IMPLEMENTATION] Secret scanning enabled — no credentials committed to repository",
        "[IMPLEMENTATION] Peer code review with security focus completed",
        "[TESTING] Security test cases written for critical functionality",
        "[TESTING] DAST scan completed in staging environment",
        "[TESTING] API security testing performed",
        "[TESTING] Penetration test completed (required for critical/external-facing applications)",
        "[DEPLOYMENT] Security review sign-off obtained before production release",
        "[DEPLOYMENT] Rollback plan tested and available",
        "[DEPLOYMENT] Security monitoring/alerting verified active post-deployment",
    ]
    for i, activity in enumerate(sdlc_activities):
        r = row + i
        # Col A: empty FFFFCC (App-ID for user to fill)
        apply_style(ws.cell(row=r, column=1), styles['input_cell'])
        # Col B: activity description text — FFFFCC
        ws.cell(row=r, column=2).value = activity
        apply_style(ws.cell(row=r, column=2), styles['input_cell'])
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Cols C-N: empty FFFFCC input
        for col_num in range(3, 15):
            apply_style(ws.cell(row=r, column=col_num), styles['input_cell'])

    row += len(sdlc_activities)

    # Add 50 empty FFFFCC rows after activity rows
    for empty_row in range(row, row + 50):
        for col_num in range(1, 15):  # 14 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'

    return ws


def build_secure_coding_standards_sheet(wb, styles, validations):
    """Build Secure Coding Standards adoption sheet."""
    ws = wb["Secure Coding Standards"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "SECURE CODING STANDARDS ADOPTION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "ASSESS SECURE CODING STANDARDS ADOPTION PER APPLICATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Standard Adopted",
        "Standard Documented?",
        "Developers Trained?",
        "Enforced via Tools?",
        "Enforced via Code Review?",
        "Last Standard Update",
        "Compliance Score (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    for i in range(3, 8):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.column_dimensions['H'].width = 18
    
    # Apply data validations
    validations['yes_no'].add('C4:F100')
    
    # Example data with formula
    example_data = [
        ["Customer Portal", "OWASP Secure Coding Practices + Java Secure Coding (Oracle)", "Yes", "Yes", "Yes", "Yes", "2024-11-15"],
        ["Internal HR System", "OWASP + Python Secure Coding", "Yes", "Yes", "Yes", "No", "2024-09-20"],
        ["Marketing Website", "OWASP Baseline", "Yes", "No", "Yes", "Yes", "2024-10-10"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])

        # Add Compliance Score formula in column H
        cell_h = ws.cell(row=row, column=8)
        cell_h.value = f'=(IF(C{row}="Yes",25,0)+IF(D{row}="Yes",25,0)+IF(E{row}="Yes",25,0)+IF(F{row}="Yes",25,0))'
        cell_h.number_format = '0"%"'
        apply_style(cell_h, styles['sample_cell'])

        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 9):  # 8 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_code_review_metrics_sheet(wb, styles, validations):
    """Build Code Review Metrics sheet."""
    ws = wb["Code Review Metrics"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "CODE REVIEW METRICS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "TRACK CODE REVIEW PROCESS AND SECURITY FOCUS"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Code Review Process Documented?",
        "Code Review Coverage (%)",
        "Security Checklist Used?",
        "Security Champion Involved?",
        "Avg Review Turnaround (days)",
        "Security Findings Count",
        "Review Compliance Score (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 22
    
    # Apply data validations
    validations['yes_no'].add('B4:B100')
    validations['yes_no'].add('D4:E100')
    
    # Example data with formula
    example_data = [
        ["Customer Portal", "Yes", 100, "Yes", "Yes", 1.5, 8],
        ["Internal HR System", "Yes", 95, "Yes", "No", 2.2, 12],
        ["Marketing Website", "Yes", 85, "No", "Yes", 1.8, 5],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])
            if col_num == 3:  # Coverage percentage
                cell.number_format = '0"%"'

        # Add Review Compliance Score formula in column H
        cell_h = ws.cell(row=row, column=8)
        cell_h.value = f'=IF(B{row}="Yes",20,0)+(C{row}*0.4)+IF(D{row}="Yes",20,0)+IF(E{row}="Yes",20,0)'
        cell_h.number_format = '0.0"%"'
        apply_style(cell_h, styles['sample_cell'])

        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 9):  # 8 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_security_tools_deployment_sheet(wb, styles, validations):
    """Build Security Tools Deployment sheet."""
    ws = wb["Security Tools Deployment"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "SECURITY TOOLS DEPLOYMENT STATUS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "TRACK SECURITY TOOL DEPLOYMENT ACROSS ORGANISATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Tool Type",
        "Tool Name/Vendor",
        "Deployment Status",
        "Applications Covered",
        "Integration Point",
        "Config Reviewed?",
        "Findings Per Month (Avg)",
        "False Positive Rate (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    # Apply data validations
    validations['tool_status'].add('C4:C100')
    validations['yes_no'].add('F4:F100')
    
    # Example data
    example_data = [
        ["SAST", "SonarQube Community", "Deployed", "15 applications", "CI/CD (GitLab)", "Yes", 125, 15],
        ["SCA", "Snyk Open Source", "Deployed", "12 applications", "CI/CD (GitLab)", "Yes", 45, 8],
        ["Secret Scanning", "TruffleHog", "Deployed", "All repositories", "Pre-commit hook", "Yes", 3, 5],
        ["DAST", "OWASP ZAP", "In Progress", "5 applications", "Test Pipeline", "No", 0, 0],
        ["Container Scanning", "Trivy", "Deployed", "8 applications", "CI/CD (Docker build)", "Yes", 22, 10],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])
            if col_num == 8:  # False positive rate
                cell.number_format = '0.0"%"'
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 9):  # 8 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_security_tools_usage_sheet(wb, styles, validations):
    """Build Security Tools Usage sheet."""
    ws = wb["Security Tools Usage"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "SECURITY TOOLS USAGE METRICS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "TRACK SECURITY TOOL USAGE PER APPLICATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "SAST Scans Per Month",
        "SCA Scans Per Month",
        "Secret Scanning Enabled?",
        "DAST Scans Per Release",
        "Avg Remediation Time (days)",
        "Tool Integration Score",
        "Usage Compliance (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    # Apply data validations
    validations['yes_no'].add('D4:D100')
    
    # Example data with formula
    example_data = [
        ["Customer Portal", 20, 20, "Yes", 2, 5.5, "Excellent"],
        ["Internal HR System", 8, 8, "Yes", 1, 12.0, "Good"],
        ["Marketing Website", 16, 16, "Yes", 4, 3.2, "Excellent"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])

        # Add Usage Compliance formula in column H
        cell_h = ws.cell(row=row, column=8)
        cell_h.value = f'=(IF(B{row}>=4,25,0)+IF(C{row}>=4,25,0)+IF(D{row}="Yes",25,0)+IF(E{row}>=1,25,0))'
        cell_h.number_format = '0"%"'
        apply_style(cell_h, styles['sample_cell'])

        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 9):  # 8 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_developer_training_sheet(wb, styles, validations):
    """Build Developer Training sheet."""
    ws = wb["Developer Training"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "DEVELOPER SECURITY TRAINING COMPLIANCE"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "TRACK DEVELOPER SECURITY TRAINING COMPLETION AND COMPLIANCE"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Developer/Team Name",
        "Initial Training Date",
        "Annual Refresher Date",
        "Language-Specific Training",
        "Security Champion Training",
        "Training Quiz Score (%)",
        "Training Overdue?",
        "Training Status",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    # Example data with formulas
    example_data = [
        ["Sandra Brunner", "2024-03-15", "2025-01-10", "Java Secure Coding (2024-06-20)", "N/A", 92],
        ["Markus Huber", "2023-05-10", "2024-01-15", "N/A", "Champion Training (2024-08-12)", 88],
        ["Petra Keller", "2024-08-22", "N/A", "Python Security (2024-10-05)", "N/A", 85],
        ["Development Team A", "85% trained", "70% current", "60% completed", "2 champions", 86],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            if col_num <= 6:
                cell.value = value
                apply_style(cell, styles['sample_cell'])
                if col_num == 6 and isinstance(value, (int, float)):  # Quiz score
                    cell.number_format = '0"%"'
            elif col_num == 7:  # Training Overdue (Column G)
                # Formula: Check if annual refresher > 365 days old
                if row == 4:  # Individual developer rows
                    cell.value = f'=IF(AND(C{row}<>"",C{row}<TODAY()-365),"Yes","No")'
                else:
                    cell.value = "VARIES"
                apply_style(cell, styles['sample_cell'])
            else:  # Training Status (Column H)
                if row == 4:
                    cell.value = f'=IF(AND(B{row}<>"",C{row}<>"",C{row}>""-365),"Current","⚠️ Overdue")'
                else:
                    cell.value = "MIXED"
                apply_style(cell, styles['sample_cell'])
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 9):  # 8 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_security_defect_management_sheet(wb, styles, validations):
    """Build Security Defect Management sheet."""
    ws = wb["Security Defect Mgmt"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "SECURITY DEFECT MANAGEMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "TRACK OPEN SECURITY DEFECTS AND REMEDIATION COMPLIANCE"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Critical Defects",
        "High Defects",
        "Medium Defects",
        "Low Defects",
        "Total Open Defects",
        "Avg Age (days)",
        "SLA Compliance (%)",
        "Security Tech Debt",
        "Monthly Trend",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    for i in range(2, 11):
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    # Example data
    example_data = [
        ["Customer Portal", 0, 1, 3, 8, 12, 15, 92, 2, "Improving"],
        ["Internal HR System", 0, 2, 5, 12, 19, 28, 78, 5, "Stable"],
        ["Marketing Website", 0, 0, 2, 4, 6, 8, 100, 0, "Improving"],
        ["Employee Portal", 1, 3, 4, 10, 18, 45, 65, 8, "Degrading"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])
            if col_num == 8:  # SLA Compliance percentage
                cell.number_format = '0"%"'
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 11):  # 10 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    ws.freeze_panes = 'A4'
    
    return ws


def build_compliance_summary_sheet(wb, styles, validations):
    """Build Summary Dashboard with TABLE 1/2/3 gold standard structure."""
    ws = wb["Summary Dashboard"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # =========================================================================
    # ROW 1: MAIN TITLE
    # =========================================================================
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value=f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border

    # =========================================================================
    # ROW 2: CONTROL REFERENCE
    # =========================================================================
    ws.merge_cells("A2:G2")
    cell = ws.cell(row=2, column=1, value=CONTROL_REF)
    cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    # =========================================================================
    # ROW 4: TABLE 1 — COMPLIANCE OVERVIEW
    # =========================================================================
    # Area configs: (display_name, sheet_name, status_col, complete_val, partial_val, nc_val, row_start, row_end)
    area_configs = [
        ("SDLC Phase Activities",   "SDLC Phase Activities",    "D", "Complete",  "In Progress",  "Not Done",     5, 75),
        ("Code Review Process",     "Code Review Metrics",      "B", "Yes",       None,           "No",           5, 54),
        ("Security Tool Deployment","Security Tools Deployment","C", "Deployed",  "In Progress",  "Not Deployed", 5, 54),
        ("Developer Training",      "Developer Training",       "H", "Current",   None,           "\u26a0\ufe0f Overdue", 5, 54),
    ]

    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row = 5
    t1_headers = ["Assessment Area", "Total Assessed", "Complete", "Partial", "Missing", "N/A", "Completion %"]
    for col_idx, h in enumerate(t1_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, (display_name, sheet_name, s_col, complete_v, partial_v, nc_v, r_start, r_end) in enumerate(area_configs):
        r = 6 + i
        ws.cell(row=r, column=1, value=display_name).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border

        c = ws.cell(row=r, column=2)
        c.value = f"=COUNTA('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end})"
        c.border = border; c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=r, column=3)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{complete_v}\")"
        c.border = border; c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=r, column=4)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{partial_v}\")" if partial_v else "=0"
        c.border = border; c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=r, column=5)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{nc_v}\")"
        c.border = border; c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=r, column=6)
        c.value = "=0"
        c.border = border; c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=r, column=7)
        c.value = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
        c.number_format = "0.0%"
        c.border = border; c.alignment = Alignment(horizontal="center")

    total_row = 6 + len(area_configs)  # row 10
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}6:{col_letter}{total_row-1})"
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border; c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=total_row, column=7)
    c.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.border = border; c.alignment = Alignment(horizontal="center")

    # =========================================================================
    # TABLE 2: KEY METRICS
    # =========================================================================
    row = total_row + 2  # row 12
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    metrics = [
        ("Total Applications Assessed",       "=COUNTA('SDLC Phase Activities'!B5:B75)"),
        ("Security Tools Deployed (Total)",    "=COUNTIF('Security Tools Deployment'!C5:C54,\"Deployed\")"),
        ("Developers Trained (Current)",       "=COUNTIF('Developer Training'!H5:H54,\"Current\")"),
        ("Next Review Due",                    ""),
        ("Assessment Owner",                   ""),
    ]

    # TABLE 2 column headers
    header_row = row + 1
    ws.cell(row=header_row, column=1, value="Metric").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=header_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=header_row, column=1).border = border
    ws.merge_cells(f"B{header_row}:G{header_row}")
    ws.cell(row=header_row, column=2, value="Value").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=header_row, column=2).fill = PatternFill("solid", fgColor="D9D9D9")
    for col in range(2, 8):
        ws.cell(row=header_row, column=col).border = border
    r = row + 2
    for label, formula in metrics:
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(row=r, column=2)
        c.value = formula if formula else ""
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # =========================================================================
    # TABLE 3: CRITICAL FINDINGS
    # =========================================================================
    # 2 buffer rows: A alone | B:G merged (mirrors TABLE 2 metric structure)
    for _ in range(2):
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:G{r}")
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = border
        r += 1
    row = r + 1  # blank gap, then TABLE 3
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: CRITICAL FINDINGS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    t3_headers = ["#", "Finding", "Count", "Severity", "Affected Area", "Recommended Action", "Owner"]
    for col_idx, h in enumerate(t3_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border

    findings = [
        ("=COUNTIF('SDLC Phase Activities'!D5:D75,\"Not Done\")",        "SDLC Requirements Activities Not Completed", "High",   "SDLC Phase Activities",     "Define security requirements for all applications"),
        ("=COUNTIF('Code Review Metrics'!B5:B54,\"No\")",                 "Apps Without Code Review Process",           "High",   "Code Review Metrics",       "Implement mandatory code review process"),
        ("=COUNTIF('Security Tools Deployment'!C5:C54,\"Not Deployed\")", "Security Tools Not Deployed",                "Medium", "Security Tools Deployment", "Deploy and configure missing security tooling"),
        ("=COUNTIF('Developer Training'!H5:H54,\"⚠️ Overdue\")", "Developers With Overdue Security Training", "High",   "Developer Training",        "Schedule security training for overdue developers"),
        ("=COUNTIF('Security Tools Usage'!D5:D54,\"No\")",                "Apps Without Secret Scanning Enabled",       "Medium", "Security Tools Usage",      "Enable secret scanning for all repositories"),
    ]

    for i, (count_f, finding_name, severity, area, action) in enumerate(findings, 1):
        r = row + i
        ws.cell(row=r, column=1, value=count_f).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=finding_name).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=count_f).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=4, value=severity).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=5, value=area).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=6, value=action).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=6).border = border
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=7).border = border

    for col, w in zip("ABCDEFG", [38, 16, 16, 16, 16, 16, 18]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"
    return ws
def create_evidence_register(wb, styles):
    """Create Evidence Register — gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["Evidence Register"]
    ws.sheet_view.showGridLines = False

    # ROW 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border_thin

    # ROW 2: Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border_thin

    # ROW 3: intentionally empty (visual separator)

    # ROW 4: Column headers (003366 navy)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    dv_type = DataValidation(
        type="list",
        formula1='"SDLC process doc,Code review report,Training record,Tool configuration,Security policy,Test report,Screenshot,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"\u2705 Verified,\u26a0\ufe0f Pending,\u274c Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    # ROW 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001",
        2: "SDLC Security Activities Assessment",
        3: "Policy Document",
        4: "SDLC security activities and phase compliance documentation",
        5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "\u2705 Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    dv_type.add(ws["C5"])
    dv_ver.add(ws["H5"])

    # ROWS 6-105: Empty data rows (FFFFCC yellow, 100 rows — MAX-002)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None  # No pre-filled IDs (MAX-001)
        dv_type.add(ws[f"C{r}"])
        dv_ver.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"
    return ws
def create_approval_sheet(wb, styles, doc_id, assessment_name, summary_sheet_name="Summary Dashboard", compliance_cell="G10"):
    """Create Approval Sign-Off — gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    if "Approval Sign-Off" in wb.sheetnames:
        del wb["Approval Sign-Off"]
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    # ROW 1: Title banner (003366)
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border_thin

    # ROW 2: Control reference (003366 italic, NO fill)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border_thin

    # ROW 3: Assessment Summary banner (4472C4)
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border_thin

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:",                  f"{doc_id} - {assessment_name}"),
        ("Assessment Period:",         ""),
        ("Overall Compliance Rating:", f"='{summary_sheet_name}'!G10"),
        ("Assessment Status:",         ""),
        ("Assessed By:",               ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border_thin  # GS-AS-011: border on label column
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for col_letter in ['B', 'C', 'D', 'E']:
            ws[f"{col_letter}{row}"].border = border_thin
        row += 1

    # Assessment Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(ws["B7"])  # Row 7 = Assessment Status

    # Approver sections
    row += 2  # row 10
    for title, color in [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_thin
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = border_thin
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for col_letter in ['B', 'C', 'D', 'E']:
                ws[f"{col_letter}{row}"].border = border_thin
            row += 1
        row += 1  # gap between sections

    # Final Decision
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border_thin
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for col_letter in ['B', 'C', 'D', 'E']:
        ws[f"{col_letter}{row}"].border = border_thin
    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # Next Review Details
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border_thin

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border_thin
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for col_letter in ['B', 'C', 'D', 'E']:
            ws[f"{col_letter}{row}"].border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    return ws
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            dv.showErrorMessage = True
            dv.showInputMessage = True
def main():
    """Main execution function."""
    logger.info("Generating ISMS A.8.25 SDLC Security Activities Assessment Workbook...")
    logger.info("=" * 70)
    
    # Create workbook and styles
    logger.info("\nCreating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    # Build all sheets
    logger.info("Building Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])
    
    logger.info("Building SDLC Phase Activities sheet...")
    build_sdlc_phase_activities_sheet(wb, styles, None)
    
    logger.info("Building Secure Coding Standards sheet...")
    build_secure_coding_standards_sheet(wb, styles, None)

    logger.info("Building Code Review Metrics sheet...")
    build_code_review_metrics_sheet(wb, styles, None)

    logger.info("️  Building Security Tools Deployment sheet...")
    build_security_tools_deployment_sheet(wb, styles, None)

    logger.info("Building Security Tools Usage sheet...")
    build_security_tools_usage_sheet(wb, styles, None)

    logger.info(" Building Developer Training sheet...")
    build_developer_training_sheet(wb, styles, None)

    logger.info(" Building Security Defect Management sheet...")
    build_security_defect_management_sheet(wb, styles, None)

    logger.info(" Building Summary Dashboard dashboard...")
    build_compliance_summary_sheet(wb, styles, None)
    
    logger.info(" Building Evidence Register...")
    create_evidence_register(wb, styles)
    
    logger.info("Building Approval Sign-Off sheet...")
    create_approval_sheet(wb, styles, "ISMS-IMP-A.8.25-26-29.S2", "SDLC Security Activities Assessment", "Summary Dashboard", "G10")
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.25-26-29.S2_SDLC_Security_Activities_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    logger.info(f"\n Saving workbook: {filename}")
    # Finalize all data validations

    for ws_name in wb.sheetnames:

        ws_temp = wb[ws_name]

        for dv in list(ws_temp.data_validations.dataValidation):

            dv.showErrorMessage = True

            dv.showInputMessage = True

    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info("\n" + "=" * 70)
    logger.info("Workbook generated successfully!")
    logger.info("=" * 70)
    logger.info(f"\nFile: {filename}")
    logger.info(f" Sheets: {len(wb.sheetnames)}")
    logger.info("\nSheet List:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        logger.info(f"  {i}. {sheet_name}")
    
    logger.info("\n" + "=" * 70)
    logger.info("NEXT STEPS:")
    logger.info("=" * 70)
    logger.info("1. Open the workbook in Excel or LibreOffice Calc")
    logger.info("2. Review the Instructions & Legend sheet")
    logger.info("3. Complete SDLC_Phase_Activities for each application")
    logger.info("4. Document secure coding standards adoption")
    logger.info("5. Track code review metrics")
    logger.info("6. Document security tool deployment and usage")
    logger.info("7. Track developer training compliance")
    logger.info("8. Monitor security defect management")
    logger.info("9. Review Compliance_Summary for overall scores")
    logger.info("10. Obtain approvals in Approval_Sign_Off")
    logger.info("\n" + "=" * 70)
    logger.info(" REFERENCE:")
    logger.info("=" * 70)
    logger.info("Implementation Guide: ISMS-IMP-A.8.25-26-29-S2")
    logger.info("Policy Reference: ISMS-POL-A.8.25-26-29-S3")
    logger.info("ISO Control: A.8.25 (Secure Development Lifecycle)")
    logger.info("=" * 70)
    


if __name__ == "__main__":
    try:
        filename = main()
        logger.info(f"\nSUCCESS: {filename} created successfully\n")
    except Exception as e:
        logger.error(f"\nERROR: {str(e)}\n")
        import traceback
        traceback.print_exc()
        exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
