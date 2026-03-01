#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.25-26-29.S1 - Security Requirements Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.25/A.8.26/A.8.29: Secure Development Framework
Assessment Domain 1 of 5: Application Security Requirements (A.8.26)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific:
- Development methodologies (Agile, Waterfall, DevOps, hybrid)
- Technology stack (languages, frameworks, platforms)
- Security tooling (SAST, DAST, SCA, IAST tools and versions)
- SDLC processes and governance structure
- Compliance requirements and risk tolerance
- Application portfolio and classification scheme
- Security requirements taxonomy and threat model frameworks
- Security architecture review processes

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.25-26-29 Secure Development Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
application security requirements specification and validation processes across
the application portfolio.

**Purpose:**
Enables systematic assessment of security requirements definition, documentation,
threat modeling, architecture review, and traceability against ISO 27001:2022
Control A.8.26 requirements, supporting evidence-based validation of security
requirements practices.

**Assessment Scope:**
- Application inventory with risk classification
- Security requirements documentation completeness
- Functional security requirements coverage
- Non-functional security requirements coverage
- Threat modeling execution and documentation
- Security architecture review completion
- Requirements-to-implementation traceability
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and security requirements standards
2. Application Inventory - Complete portfolio with risk classifications
3. Requirements Documentation Status - Documentation completeness per application
4. Threat Modeling Status - Threat model execution and review status
5. Architecture Review Status - Security architecture review completion
6. Traceability Matrix Status - Requirements traceability compliance
7. Summary Dashboard - Overall security requirements compliance scores
8. Gap Analysis - Non-compliant applications and remediation requirements
9. Evidence Register - Audit evidence tracking and documentation
10. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with application risk classification dropdown lists
- Conditional formatting for requirements completeness status
- Automated gap identification for missing requirements documentation
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with threat modeling tools and security architecture reviews

**Integration:**
This assessment feeds into the A.8.25-26-29.5 Secure Development Compliance
Dashboard, which consolidates data from all five assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl>=3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a825_26_29_1_security_requirements.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a825_26_29_1_security_requirements.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a825_26_29_1_security_requirements.py --date 20250124

Output:
    File: ISMS_A_8_25_26_29_1_Security_Requirements_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize application risk classification criteria
    2. Inventory all applications in your portfolio (internal, vendor, cloud)
    3. Complete requirements documentation status for each application
    4. Validate threat modeling completion for high-risk applications
    5. Review security architecture review status
    6. Verify requirements traceability for critical applications
    7. Conduct gap analysis for applications with missing requirements
    8. Define remediation actions with application owners and timelines
    9. Collect and link audit evidence (requirements docs, threat models, reviews)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.25-26-29.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.25/A.8.26/A.8.29
Assessment Domain:    1 of 5 (Application Security Requirements - A.8.26)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [YYYY-MM-DD]
Last Modified:        [YYYY-MM-DD]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.25-26-29-S1: Executive Control Alignment (Governance)
    - ISMS-POL-A.8.25-26-29-S2: Security Requirements (A.8.26)
    - ISMS-POL-A.8.25-26-29-S3: Secure Development Lifecycle (A.8.25)
    - ISMS-POL-A.8.25-26-29-S4: Security Testing (A.8.29)
    - ISMS-POL-A.8.25-26-29-S5: Assessment Evidence Framework
    - ISMS-IMP-A.8.25-26-29-S1: Security Requirements Process Implementation Guide
    - ISMS-IMP-A.8.25-26-29.S5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a825_26_29_1_security_requirements.py (this script)
    - generate_a825_26_29_2_sdlc_security_activities.py
    - generate_a825_26_29_3_security_testing_results.py
    - generate_a825_26_29_4_vulnerability_remediation.py
    - generate_a825_26_29_5_compliance_dashboard.py
    - normalize_assessment_files_a825_26_29.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [YYYY-MM-DD]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.25-26-29-S1 spec
    - Supports comprehensive security requirements evaluation
    - Integrated with A.8.25-26-29.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**SDLC Methodology Neutrality:**
This assessment framework is methodology-agnostic and must work across:
- Waterfall: Traditional phase-gate development with upfront requirements
- Agile: Iterative sprint-based development with evolving requirements
- DevOps/DevSecOps: Continuous integration/deployment with security automation
- Hybrid: Mixed methodology approaches
Customize dropdown values and assessment criteria to match YOUR methodology.

**Technology Stack Agnosticism:**
Framework must remain vendor-neutral and technology-agnostic:
- Programming languages: Java, Python, C#, JavaScript, Go, Ruby, etc.
- Frameworks: Spring, Django, .NET, React, Angular, Vue, etc.
- Platforms: Web, mobile, desktop, embedded, cloud-native, SaaS, etc.
Do NOT hardcode technology-specific assumptions in assessment criteria.

**Security Tool Diversity:**
Assessment must accommodate various security requirements and threat modeling tools:
- Threat Modeling: Microsoft Threat Modeling Tool, OWASP Threat Dragon, IriusRisk
- Requirements Management: Jira, Azure DevOps, Confluence, custom systems
- Architecture Tools: Enterprise Architect, draw.io, Lucidchart, C4 Model
Use generic terms; avoid vendor lock-in in assessment structure.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Complete application inventory with risk classifications
- Documented security requirements for each application
- Threat models for high-risk applications
- Security architecture review records
- Requirements traceability matrices
- Evidence of requirements validation and approval

**Data Protection:**
Assessment workbooks contain sensitive application information including:
- Application portfolio and architecture details
- Risk classifications and threat assessments
- Security requirements and control implementations
- Security gaps and vulnerabilities
- Internal system names and data flows
Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Complete reassessment of all applications
- Semi-annually: Update risk classifications based on threat landscape
- Annually: Review and update security requirements taxonomy
- Ad-hoc: When new applications are deployed or architectures change

**Quality Assurance:**
Have Application Security team, Security Architects, and Development Leads
validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
Security requirements assessment supports regulatory compliance:
- Payment processing: PCI DSS v4.0.1 requirement 6.3 (secure development)
- Healthcare: HIPAA security requirements for applications handling PHI
- Finance: Regional banking requirements for secure application development
- Privacy: GDPR/FADP privacy-by-design and data protection requirements
Customize assessment criteria to include regulatory-specific requirements
applicable to your organisation and industry.

**Integration with Related Controls:**
This assessment integrates with:
- A.8.4 (Access to Source Code): Requirements for repository access controls
- A.8.28 (Secure Coding): Requirements inform secure coding standards
- A.8.31 (Environment Separation): Requirements for environment isolation
- A.8.32 (Change Management): Requirements traceability in change process
- A.8.25 (Secure Development Lifecycle): Requirements feed into SDLC phases
- A.8.29 (Security Testing): Requirements define test cases and acceptance criteria

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

from datetime import datetime, timedelta
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.25-26-29.S1"
CONTROL_ID   = "A.8.25-26-29"
CONTROL_NAME = "Secure Development"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
WORKBOOK_NAME = "Security Requirements Assessment"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Sheet structure from ISMS-IMP-A.8.25-26-29-S1 Section 6.2
    sheets = [
        "Instructions & Legend",
        "Application Inventory",
        "Requirements Documentation",
        "Threat Modelling",
        "Architecture Review",
        "Traceability Matrix",
        "Evidence Register",
        "Gap Analysis",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "sample_cell": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True),
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True),
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True),
        },
        "percentage_good": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "percentage_fair": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "percentage_poor": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name if hasattr(style_dict["font"], 'name') else "Calibri",
            size=style_dict["font"].size if hasattr(style_dict["font"], 'size') else 10,
            bold=style_dict["font"].bold if hasattr(style_dict["font"], 'bold') else False,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type if hasattr(style_dict["fill"], 'fill_type') else "solid"
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal if hasattr(style_dict["alignment"], 'horizontal') else "left",
            vertical=style_dict["alignment"].vertical if hasattr(style_dict["alignment"], 'vertical') else "center",
            wrap_text=style_dict["alignment"].wrap_text if hasattr(style_dict["alignment"], 'wrap_text') else False
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create FRESH data validation objects registered with ws.
    Call once per sheet builder with that sheet's ws to avoid cross-sheet DV contamination.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial"',
            allow_blank=False
        ),
        'yes_no_planned': DataValidation(
            type="list",
            formula1='"Yes,No,Planned"',
            allow_blank=False
        ),
        'yes_no_scheduled': DataValidation(
            type="list",
            formula1='"Yes,No,Scheduled"',
            allow_blank=False
        ),
        'risk_classification': DataValidation(
            type="list",
            formula1='"High Risk,Medium Risk,Low Risk"',
            allow_blank=False
        ),
        'approval_status': DataValidation(
            type="list",
            formula1='"Approved,Pending,Not Approved"',
            allow_blank=False
        ),
        'threat_methodology': DataValidation(
            type="list",
            formula1='"STRIDE,PASTA,LINDDUN,Other,N/A"',
            allow_blank=False
        ),
        'evidence_status': DataValidation(
            type="list",
            formula1='"Current,Outdated,Missing"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1='"Compliant,⚠️ Partial Compliance,Non-Compliant"',
            allow_blank=False
        ),
    }
    
    # Add validations to worksheet
    for key, dv in validations.items():
        ws.add_data_validation(dv)
    
    return validations


# ============================================================================
# SECTION 3: SHEET BUILDERS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab for applicable systems/services.', '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Check all applicable items in the Compliance Checklist for each section.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def build_application_inventory_sheet(wb, styles):
    """Build Application Inventory sheet."""
    ws = wb["Application Inventory"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet

    # Title
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "APPLICATION INVENTORY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "COMPLETE INVENTORY OF ALL APPLICATIONS REQUIRING SECURITY REQUIREMENTS ASSESSMENT"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Application Name",
        "Application ID",
        "Application Owner",
        "Business Criticality",
        "Risk Classification",
        "Technology Stack",
        "Deployment Model",
        "Regulatory Scope",
        "Internet Facing",
        "Last Assessment Date",
    ]

    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Column widths
    ws.column_dimensions['A'].width = 25  # Application Name
    ws.column_dimensions['B'].width = 15  # Application ID
    ws.column_dimensions['C'].width = 20  # Owner
    ws.column_dimensions['D'].width = 18  # Business Criticality
    ws.column_dimensions['E'].width = 18  # Risk Classification
    ws.column_dimensions['F'].width = 22  # Technology Stack
    ws.column_dimensions['G'].width = 18  # Deployment Model
    ws.column_dimensions['H'].width = 20  # Regulatory Scope
    ws.column_dimensions['I'].width = 15  # Internet Facing
    ws.column_dimensions['J'].width = 18  # Last Assessment

    # Apply data validations
    validations['risk_classification'].add('E4:E100')
    validations['yes_no'].add('I4:I100')

    # Example data rows
    example_data = [
        ["Customer Portal", "APP-001", "Anna Müller", "High", "High Risk", "Java/Spring Boot", "AWS Cloud", "GDPR, PCI DSS v4.0.1", "Yes", "2025-01-15"],
        ["Internal HR System", "APP-002", "Thomas Meier", "Medium", "Medium Risk", "Python/Django", "On-Premises", "GDPR", "No", "2024-12-10"],
        ["Marketing Website", "APP-003", "Sandra Brunner", "Medium", "Medium Risk", "Static HTML/CSS", "CDN (Cloudflare)", "None", "Yes", "2025-01-05"],
        ["Employee Portal", "APP-004", "Markus Huber", "Medium", "Medium Risk", "ASP.NET Core", "Azure Cloud", "GDPR", "No", "2024-11-22"],
        ["Finance Dashboard", "APP-005", "Petra Keller", "High", "High Risk", "React + Node.js", "AWS Cloud", "SOX, GDPR", "No", "2025-01-08"],
    ]

    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 11):  # 10 columns (A-J)
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    # Freeze panes
    ws.freeze_panes = 'A4'

    return ws


def build_requirements_documentation_sheet(wb, styles):
    """Build Requirements Documentation Status sheet."""
    ws = wb["Requirements Documentation"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "SECURITY REQUIREMENTS DOCUMENTATION ASSESSMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "ASSESS COMPLETENESS OF SECURITY REQUIREMENTS DOCUMENTATION PER APPLICATION"
    apply_style(cell, styles['subheader'])

    # Column headers — App-ID in col A, Application Name in col B, remaining shift right
    headers = [
        "App-ID",
        "Application Name",
        "Requirements Document Exists?",
        "Document Location/Link",
        "Last Updated",
        "Requirements Approved?",
        "Approver Name",
        "Approval Date",
        "Functional Requirements (%)",
        "Non-Functional Requirements (%)",
        "Data Protection Requirements (%)",
        "Completeness Score (%)",
    ]

    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Column widths
    ws.column_dimensions['A'].width = 12   # App-ID
    ws.column_dimensions['B'].width = 25   # Application Name
    ws.column_dimensions['C'].width = 18   # Requirements Document Exists?
    ws.column_dimensions['D'].width = 40   # Document Location/Link
    ws.column_dimensions['E'].width = 15   # Last Updated
    ws.column_dimensions['F'].width = 18   # Requirements Approved?
    ws.column_dimensions['G'].width = 20   # Approver Name
    ws.column_dimensions['H'].width = 15   # Approval Date
    ws.column_dimensions['I'].width = 20   # Functional Requirements (%)
    ws.column_dimensions['J'].width = 22   # Non-Functional Requirements (%)
    ws.column_dimensions['K'].width = 22   # Data Protection Requirements (%)
    ws.column_dimensions['L'].width = 18   # Completeness Score (%)

    # Apply data validations (shifted right by 1 due to new App-ID col A)
    validations['yes_no_partial'].add('C4:C100')
    validations['yes_no'].add('F4:F100')

    # Example data: App-ID in col A, then remaining fields
    example_data = [
        ["APP-001", "Customer Portal", "Yes", "/docs/security-requirements/APP-001-sec-req.pdf", "2025-01-10", "Yes", "Anna Müller", "2025-01-12", 95, 90, 100],
    ]

    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])

        # Add Completeness Score formula in column L (shifted from K)
        cell_l = ws.cell(row=row, column=12)
        cell_l.value = f'=IF(AND(I{row}<>"",J{row}<>"",K{row}<>""),(I{row}+J{row}+K{row})/3,"")'
        cell_l.number_format = '0.0"%"'
        apply_style(cell_l, styles['sample_cell'])

        row += 1

    # ISO 27002:2022 A.8.26: Supplier/third-party security requirements (explicit requirement)
    # Col A = empty FFFFCC (App-ID for user to fill), Col B = requirement description FFFFCC
    supplier_items = [
        "[SUPPLIER] Externally acquired software subject to same security requirements as internally developed",
        "[SUPPLIER] Vendor/supplier security requirements specified in contracts and SoW",
        "[SUPPLIER] SaaS/PaaS applications assessed against organisational security requirements before adoption",
        "[SUPPLIER] Open source components assessed for security before inclusion in approved component list",
    ]
    for i, item in enumerate(supplier_items):
        r = row + i
        # Col A: empty FFFFCC input (App-ID)
        apply_style(ws.cell(row=r, column=1), styles['input_cell'])
        # Col B: requirement description text — FFFFCC
        ws.cell(row=r, column=2).value = item
        apply_style(ws.cell(row=r, column=2), styles['input_cell'])
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Cols C-L: empty FFFFCC input
        for col_num in range(3, 13):
            apply_style(ws.cell(row=r, column=col_num), styles['input_cell'])

    row += len(supplier_items)

    # Add 50 empty FFFFCC rows after supplier requirement rows
    for empty_row in range(row, row + 50):
        for col_num in range(1, 13):  # 12 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    # Freeze panes
    ws.freeze_panes = 'A4'

    return ws


def build_threat_modeling_sheet(wb, styles):
    """Build Threat Modeling Status sheet."""
    ws = wb["Threat Modelling"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "THREAT MODELING STATUS ASSESSMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "ASSESS THREAT MODELING COMPLETION AND QUALITY PER APPLICATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Threat Model Exists?",
        "Methodology Used",
        "Threat Model Location/Link",
        "Last Updated",
        "DFD Created?",
        "Threats Documented?",
        "Mitigations Defined?",
        "Approval Status",
        "Approver Name",
        "Completeness Score (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 18
    
    # Apply data validations
    validations['yes_no_planned'].add('B4:B100')
    validations['threat_methodology'].add('C4:C100')
    validations['yes_no'].add('F4:F100')
    validations['yes_no'].add('G4:G100')
    validations['yes_no'].add('H4:H100')
    validations['approval_status'].add('I4:I100')
    
    # Example data with formulas
    example_data = [
        ["Customer Portal", "Yes", "STRIDE", "/docs/threat-models/APP-001-threat-model.pdf", "2025-01-08", "Yes", "Yes", "Yes", "Approved", "Anna Müller"],
        ["Internal HR System", "Yes", "STRIDE", "/docs/threat-models/APP-002-threat-model.pdf", "2024-11-18", "Yes", "Yes", "Yes", "Approved", "Thomas Meier"],
        ["Marketing Website", "No", "N/A", "", "", "No", "No", "No", "Not Approved", ""],
        ["Employee Portal", "Yes", "STRIDE", "/docs/threat-models/APP-004-threat-model.pdf", "2024-11-12", "Yes", "Yes", "Partial", "Pending", "Markus Huber"],
        ["Finance Dashboard", "Yes", "STRIDE", "/docs/threat-models/APP-005-threat-model.pdf", "2025-01-03", "Yes", "Yes", "Yes", "Approved", "Petra Keller"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])

        # Add Completeness Score formula in column K
        cell_k = ws.cell(row=row, column=11)
        cell_k.value = f'=IF(B{row}="Yes",(IF(F{row}="Yes",25,0)+IF(G{row}="Yes",25,0)+IF(H{row}="Yes",25,0)+IF(I{row}="Approved",25,0)),"N/A")'
        cell_k.number_format = '0"%"'
        apply_style(cell_k, styles['sample_cell'])

        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 12):  # 11 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    return ws


def build_architecture_review_sheet(wb, styles):
    """Build Architecture Review Status sheet."""
    ws = wb["Architecture Review"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "SECURITY ARCHITECTURE REVIEW STATUS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "ASSESS SECURITY ARCHITECTURE REVIEW COMPLETION AND FINDINGS PER APPLICATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Review Conducted?",
        "Review Date",
        "Review Report Location/Link",
        "Reviewers",
        "Critical Findings",
        "High Findings",
        "Medium Findings",
        "Low Findings",
        "Open Findings",
        "Approval Status",
        "Approval Date",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    
    # Apply data validations
    validations['yes_no_scheduled'].add('B4:B100')
    validations['approval_status'].add('K4:K100')
    
    # Example data
    example_data = [
        ["Customer Portal", "Yes", "2025-01-09", "/docs/arch-reviews/APP-001-arch-review.pdf", "Security Team, Tech Lead", 0, 0, 2, 3, 0, "Approved", "2025-01-11"],
        ["Internal HR System", "Yes", "2024-11-19", "/docs/arch-reviews/APP-002-arch-review.pdf", "Security Architect", 0, 1, 3, 2, 0, "Approved", "2024-11-22"],
        ["Marketing Website", "No", "", "", "", 0, 0, 0, 0, 0, "Not Approved", ""],
        ["Employee Portal", "Yes", "2024-11-14", "/docs/arch-reviews/APP-004-arch-review.pdf", "Security Team", 0, 2, 1, 1, 2, "Pending", ""],
        ["Finance Dashboard", "Yes", "2025-01-04", "/docs/arch-reviews/APP-005-arch-review.pdf", "Security Architect, CTO", 0, 0, 1, 2, 0, "Approved", "2025-01-06"],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 13):  # 12 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    return ws


def build_traceability_sheet(wb, styles):
    """Build Traceability Matrix Status sheet."""
    ws = wb["Traceability Matrix"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet
    
    # Title
    ws.merge_cells('A1:I1')
    cell = ws['A1']
    cell.value = "REQUIREMENTS TRACEABILITY MATRIX STATUS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35
    
    # Subtitle
    ws.merge_cells('A2:I2')
    cell = ws['A2']
    cell.value = "ASSESS REQUIREMENTS TRACEABILITY COMPLETENESS PER APPLICATION"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name",
        "Traceability Matrix Exists?",
        "Matrix Location/Link",
        "Last Updated",
        "Requirements → Design Traced?",
        "Requirements → Code Traced?",
        "Requirements → Tests Traced?",
        "Traceability Coverage (%)",
        "Quality Score (%)",
    ]
    
    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    
    # Apply data validations
    validations['yes_no'].add('B4:B100')
    validations['yes_no_partial'].add('E4:E100')
    validations['yes_no_partial'].add('F4:F100')
    validations['yes_no_partial'].add('G4:G100')
    
    # Example data with formulas
    example_data = [
        ["Customer Portal", "Yes", "/docs/traceability/APP-001-traceability.xlsx", "2025-01-14", "Yes", "Yes", "Yes", 95],
        ["Internal HR System", "Yes", "/docs/traceability/APP-002-traceability.xlsx", "2024-12-08", "Yes", "Yes", "Partial", 85],
        ["Marketing Website", "No", "", "", "No", "No", "No", 0],
        ["Employee Portal", "Yes", "/docs/traceability/APP-004-traceability.xlsx", "2024-11-20", "Yes", "Partial", "Partial", 75],
        ["Finance Dashboard", "Yes", "/docs/traceability/APP-005-traceability.xlsx", "2025-01-07", "Yes", "Yes", "Yes", 90],
    ]
    
    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])
            if col_num == 8:  # Coverage percentage
                cell.number_format = '0"%"'

        # Add Quality Score formula in column I
        cell_i = ws.cell(row=row, column=9)
        cell_i.value = f'=IF(B{row}="Yes",(IF(E{row}="Yes",33,IF(E{row}="Partial",16.5,0))+IF(F{row}="Yes",33,IF(F{row}="Partial",16.5,0))+IF(G{row}="Yes",34,IF(G{row}="Partial",17,0))),"N/A")'
        cell_i.number_format = '0.0"%"'
        apply_style(cell_i, styles['sample_cell'])

        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 10):  # 9 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    return ws


def build_compliance_summary_sheet(wb, styles):
    """Build Summary Dashboard with TABLE 1/2/3 gold standard structure."""
    ws = wb["Summary Dashboard"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # =========================================================================
    # ROW 1: MAIN TITLE
    # =========================================================================
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value=f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border

    # =========================================================================
    # ROW 2: CONTROL REFERENCE
    # =========================================================================
    ws.merge_cells("A2:G2")
    cell = ws.cell(row=2, column=1, value=CONTROL_REF)
    cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    # =========================================================================
    # ROW 4: TABLE 1 — COMPLIANCE OVERVIEW
    # =========================================================================
    # Area configs: (display_name, sheet_name, status_col, complete_val, partial_val, nc_val, row_start, row_end)
    area_configs = [
        ("Security Requirements Doc",  "Requirements Documentation", "C", "Yes",       "Partial",   "No",  5, 58),
        ("Threat Modelling",            "Threat Modelling",           "I", "Approved",  "Pending",   "Not Approved", 5, 54),
        ("Architecture Review",         "Architecture Review",        "K", "Approved",  "Pending",   "Not Approved", 5, 54),
        ("Traceability Matrix",         "Traceability Matrix",        "F", "Yes",       "Partial",   "No",  5, 54),
    ]

    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row = 5
    t1_headers = ["Assessment Area", "Total Assessed", "Complete", "Partial", "Missing", "N/A", "Completion %"]
    for col_idx, h in enumerate(t1_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # TABLE 1 data rows
    for i, (display_name, sheet_name, s_col, complete_v, partial_v, nc_v, r_start, r_end) in enumerate(area_configs):
        r = 6 + i
        ws.cell(row=r, column=1, value=display_name).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border

        # B: Total (COUNTA)
        c = ws.cell(row=r, column=2)
        c.value = f"=COUNTA('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end})"
        c.border = border; c.alignment = Alignment(horizontal="center")

        # C: Complete
        c = ws.cell(row=r, column=3)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{complete_v}\")"
        c.border = border; c.alignment = Alignment(horizontal="center")

        # D: Partial
        c = ws.cell(row=r, column=4)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{partial_v}\")" if partial_v else "=0"
        c.border = border; c.alignment = Alignment(horizontal="center")

        # E: Missing (Non-Compliant)
        c = ws.cell(row=r, column=5)
        c.value = f"=COUNTIF('{sheet_name}'!{s_col}{r_start}:{s_col}{r_end},\"{nc_v}\")"
        c.border = border; c.alignment = Alignment(horizontal="center")

        # F: N/A (always 0 for these sheets)
        c = ws.cell(row=r, column=6)
        c.value = "=0"
        c.border = border; c.alignment = Alignment(horizontal="center")

        # G: Completion %
        c = ws.cell(row=r, column=7)
        c.value = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
        c.number_format = "0.0%"
        c.border = border; c.alignment = Alignment(horizontal="center")

    # TOTAL row
    total_row = 6 + len(area_configs)  # row 10
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border = border
    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}6:{col_letter}{total_row-1})"
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border; c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=total_row, column=7)
    c.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.border = border; c.alignment = Alignment(horizontal="center")

    # =========================================================================
    # TABLE 2: KEY METRICS
    # =========================================================================
    row = total_row + 2  # row 12
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    metrics = [
        ("Total Applications Assessed",   "=COUNTA('Application Inventory'!A5:A54)"),
        ("High Risk Applications",         "=COUNTIF('Application Inventory'!E5:E54,\"High Risk\")"),
        ("Approved Threat Models",         "=COUNTIF('Threat Modelling'!I5:I54,\"Approved\")"),
        ("Next Review Due",                ""),
        ("Assessment Owner",               ""),
    ]

    # TABLE 2 column headers
    header_row = row + 1
    ws.cell(row=header_row, column=1, value="Metric").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=header_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=header_row, column=1).border = border
    ws.merge_cells(f"B{header_row}:G{header_row}")
    ws.cell(row=header_row, column=2, value="Value").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=header_row, column=2).fill = PatternFill("solid", fgColor="D9D9D9")
    for col in range(2, 8):
        ws.cell(row=header_row, column=col).border = border
    r = row + 2
    for label, formula in metrics:
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(row=r, column=2)
        c.value = formula if formula else ""
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # =========================================================================
    # TABLE 3: CRITICAL FINDINGS
    # =========================================================================
    # 2 buffer rows: A alone | B:G merged (mirrors TABLE 2 metric structure)
    for _ in range(2):
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:G{r}")
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = border
        r += 1
    row = r + 1  # blank gap, then TABLE 3
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: CRITICAL FINDINGS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    t3_headers = ["#", "Finding", "Count", "Severity", "Affected Area", "Recommended Action", "Owner"]
    for col_idx, h in enumerate(t3_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border

    findings = [
        ("=COUNTIF('Requirements Documentation'!C5:C58,\"No\")",  "Apps Without Security Requirements Doc",       "High",     "Requirements Documentation",  "Create security requirements documentation"),
        ("=COUNTIF('Threat Modelling'!B5:B54,\"No\")",             "Apps Without Threat Model",                    "High",     "Threat Modelling",            "Conduct STRIDE threat modelling session"),
        ("=COUNTIF('Architecture Review'!B5:B54,\"No\")",          "Architecture Reviews Not Conducted",           "Medium",   "Architecture Review",         "Schedule security architecture review"),
        ("=SUM('Architecture Review'!F5:F54)",                     "Critical Architecture Findings (Total)",       "Critical", "Architecture Review",         "Immediate remediation required"),
        ("=SUM('Architecture Review'!J5:J54)",                     "Open Architecture Findings (Total)",           "High",     "Architecture Review",         "Track and remediate open findings"),
    ]

    for i, (count_f, finding_name, severity, area, action) in enumerate(findings, 1):
        r = row + i
        ws.cell(row=r, column=1, value=count_f).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=finding_name).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=count_f).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=4, value=severity).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=5, value=area).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=6, value=action).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=6).border = border
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=7).border = border

    # Column widths
    for col, w in zip("ABCDEFG", [38, 16, 16, 16, 16, 16, 18]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"
    return ws
def build_gap_analysis_sheet(wb, styles):
    """Build Gap Analysis sheet for tracking non-compliant applications and remediation."""
    ws = wb["Gap Analysis"]
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)  # FML-DV-001: fresh DV objects per sheet

    # Title
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS & REMEDIATION TRACKING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "TRACK SECURITY REQUIREMENTS GAPS AND REMEDIATION ACTIONS PER APPLICATION"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Gap ID",
        "Application Name",
        "Gap Category",
        "Gap Description",
        "Risk Level",
        "Remediation Action",
        "Owner",
        "Target Date",
        "Status",
        "Notes",
    ]

    row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 30

    # Apply data validations
    validations['risk_classification'].add('E4:E100')

    # Example gap data
    example_data = [
        ["GAP-001", "Marketing Website", "Missing Requirements", "No security requirements document exists", "High Risk", "Create security requirements document", "Sandra Brunner", "2025-03-15", "Open", "Low-risk app but needs baseline requirements"],
        ["GAP-002", "Marketing Website", "Missing Threat Model", "No threat model conducted", "Medium Risk", "Conduct STRIDE threat modeling session", "Security Team", "2025-03-30", "Open", "Schedule with app team"],
        ["GAP-003", "Marketing Website", "Missing Architecture Review", "No security architecture review performed", "Medium Risk", "Schedule architecture review", "Security Architect", "2025-04-15", "Open", "Depends on GAP-001 completion"],
        ["GAP-004", "Employee Portal", "Open Findings", "2 high-severity findings from architecture review remain open", "High Risk", "Remediate authentication bypass and input validation issues", "Markus Huber", "2025-02-28", "In Progress", "Dev team working on fixes"],
        ["GAP-005", "Internal HR System", "Partial Traceability", "Requirements to tests traceability only 85% complete", "Low Risk", "Complete traceability matrix for remaining requirements", "Dev Lead", "2025-02-15", "In Progress", "15% remaining"],
    ]

    row = 4
    for row_data in example_data[:1]:  # MAX-006: One F2F2F2 sample row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['sample_cell'])

            # Color code status
            if col_num == 9:  # Status column
                if value == "Open":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == "In Progress":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif value == "Closed":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            # Color code risk level
            if col_num == 5:  # Risk Level column
                if "High" in value:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif "Medium" in value:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "Low" in value:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

    # DS-005 fix: Add empty rows to reach minimum 20 data rows
    for empty_row in range(row, 55):  # 55 to create 51 total rows (1 sample + 50 empty)
        for col_num in range(1, 11):  # 10 columns
            cell = ws.cell(row=empty_row, column=col_num)
            apply_style(cell, styles['input_cell'])

    # Freeze panes
    ws.freeze_panes = 'A4'

    return ws


def create_evidence_register(wb, styles):
    """Create Evidence Register — gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["Evidence Register"]
    ws.sheet_view.showGridLines = False

    # ROW 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border_thin

    # ROW 2: Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border_thin

    # ROW 3: intentionally empty (visual separator)

    # ROW 4: Column headers (003366 navy)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    dv_type = DataValidation(
        type="list",
        formula1='"Security requirements doc,Code review report,Test report,Vulnerability scan,Penetration test,Remediation record,SDLC process doc,Screenshot,Other"',
        allow_blank=True,
    )
    dv_ver = DataValidation(
        type="list",
        formula1='"\u2705 Verified,\u26a0\ufe0f Pending,\u274c Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_ver)

    # ROW 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001",
        2: "Security Requirements Assessment",
        3: "Policy Document",
        4: "Security requirements document for Customer Portal application",
        5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026",
        7: "Assessor Name",
        8: "\u2705 Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    dv_type.add(ws["C5"])
    dv_ver.add(ws["H5"])

    # ROWS 6-105: Empty data rows (FFFFCC yellow, 100 rows — MAX-002)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None  # No pre-filled IDs (MAX-001)
        dv_type.add(ws[f"C{r}"])
        dv_ver.add(ws[f"H{r}"])

    ws.freeze_panes = "A5"
    return ws
def create_approval_sheet(wb, styles, doc_id, assessment_name, summary_sheet_name="Summary Dashboard", compliance_cell="G10"):
    """Create Approval Sign-Off — gold standard (STANDARD-SCR-COMMON-SHEETS.md)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    if "Approval Sign-Off" in wb.sheetnames:
        del wb["Approval Sign-Off"]
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    # ROW 1: Title banner (003366)
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border_thin

    # ROW 2: Control reference (003366 italic, NO fill)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border_thin

    # ROW 3: Assessment Summary banner (4472C4)
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border_thin

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:",                  f"{doc_id} - {assessment_name}"),
        ("Assessment Period:",         ""),
        ("Overall Compliance Rating:", f"='{summary_sheet_name}'!G10"),
        ("Assessment Status:",         ""),
        ("Assessed By:",               ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border_thin  # GS-AS-011: border on label column
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for col_letter in ['B', 'C', 'D', 'E']:
            ws[f"{col_letter}{row}"].border = border_thin
        row += 1

    # Assessment Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(ws["B7"])  # Row 7 = Assessment Status

    # Approver sections
    row += 2  # row 10
    for title, color in [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border_thin
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = border_thin
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for col_letter in ['B', 'C', 'D', 'E']:
                ws[f"{col_letter}{row}"].border = border_thin
            row += 1
        row += 1  # gap between sections

    # Final Decision
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border_thin
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for col_letter in ['B', 'C', 'D', 'E']:
        ws[f"{col_letter}{row}"].border = border_thin
    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # Next Review Details
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border_thin

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border_thin
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for col_letter in ['B', 'C', 'D', 'E']:
            ws[f"{col_letter}{row}"].border = border_thin
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    return ws
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            dv.showErrorMessage = True
            dv.showInputMessage = True
def main():
    """Main execution function."""
    logger.info("Generating ISMS A.8.26 Security Requirements Assessment Workbook...")
    logger.info("=" * 70)
    
    # Create workbook and styles
    logger.info("\nCreating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    # Build all sheets
    logger.info("Building Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])
    
    logger.info("Building Application Inventory sheet...")
    build_application_inventory_sheet(wb, styles)
    
    logger.info("Building Requirements Documentation Status sheet...")
    build_requirements_documentation_sheet(wb, styles)
    
    logger.info(" Building Threat Modeling Status sheet...")
    build_threat_modeling_sheet(wb, styles)
    
    logger.info("️  Building Architecture Review Status sheet...")
    build_architecture_review_sheet(wb, styles)
    
    logger.info(" Building Traceability Matrix Status sheet...")
    build_traceability_sheet(wb, styles)
    
    logger.info(" Building Summary Dashboard dashboard...")
    build_compliance_summary_sheet(wb, styles)

    logger.info("Building Gap Analysis sheet...")
    build_gap_analysis_sheet(wb, styles)

    logger.info(" Building Evidence Register...")
    create_evidence_register(wb, styles)
    
    logger.info("Building Approval Sign-Off sheet...")
    create_approval_sheet(wb, styles, "ISMS-IMP-A.8.25-26-29.S1", "Security Requirements Assessment", "Summary Dashboard", "G10")
    
    # Save workbook
    filename = OUTPUT_FILENAME
    logger.info(f"\n Saving workbook: {filename}")
    # Finalize all data validations

    for ws_name in wb.sheetnames:

        ws_temp = wb[ws_name]

        for dv in list(ws_temp.data_validations.dataValidation):

            dv.showErrorMessage = True

            dv.showInputMessage = True

    output_path = _wkbk_dir / OUTPUT_FILENAME
    finalize_validations(wb)
    wb.save(output_path)
    logger.info("\n" + "=" * 70)
    logger.info("Workbook generated successfully!")
    logger.info("=" * 70)
    logger.info(f"\nFile: {filename}")
    logger.info(f" Sheets: {len(wb.sheetnames)}")
    logger.info("\nSheet List:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        logger.info(f"  {i}. {sheet_name}")
    
    logger.info("\n" + "=" * 70)
    logger.info("NEXT STEPS:")
    logger.info("=" * 70)
    logger.info("1. Open the workbook in Excel or LibreOffice Calc")
    logger.info("2. Review the Instructions & Legend sheet for guidance")
    logger.info("3. Complete the Application_Inventory sheet with your applications")
    logger.info("4. For each application, complete:")
    logger.info("   - Requirements_Documentation_Status")
    logger.info("   - Threat_Modeling_Status")
    logger.info("   - Architecture_Review_Status")
    logger.info("   - Traceability_Matrix_Status")
    logger.info("5. Review the Compliance_Summary dashboard for overall scores")
    logger.info("6. Document evidence in the Evidence_Register")
    logger.info("7. Obtain approvals in the Approval_Sign_Off sheet")
    logger.info("\n" + "=" * 70)
    logger.info(" REFERENCE:")
    logger.info("=" * 70)
    logger.info("Implementation Guide: ISMS-IMP-A.8.25-26-29-S1")
    logger.info("Policy Reference: ISMS-POL-A.8.25-26-29-S2")
    logger.info("ISO Control: A.8.26 (Application Security Requirements)")
    logger.info("=" * 70)
    


if __name__ == "__main__":
    try:
        filename = main()
        logger.info(f"\nSUCCESS: {filename} created successfully\n")
    except Exception as e:
        logger.error(f"\nERROR: {str(e)}\n")
        import traceback
        traceback.print_exc()
        exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
