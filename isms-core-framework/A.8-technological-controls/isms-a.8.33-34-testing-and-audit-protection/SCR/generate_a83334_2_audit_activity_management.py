#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.33-34.2 - Audit Activity Management Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.33-34: Testing and Audit Protection
Assessment Domain 2 of 2: Audit Activity Management Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific testing and audit protection infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Test data classification and anonymisation requirements (match your data governance)
2. Audit activity categories and access scope definitions
3. Test data lifecycle and approved retention periods
4. Audit tool access control requirements and approval workflow
5. Audit log protection and integrity assurance requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.33-34 Testing and Audit Protection Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
testing and audit protection controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Audit Activity Management Assessment under ISO 27001:2022 Controls A.8.33 and A.8.34. Supports evidence-based evaluation of test data protection compliance and audit activity management effectiveness.

**Assessment Scope:**
- Test data inventory completeness and protection requirement coverage
- Data anonymisation and masking adequacy for test environments
- Audit activity authorisation and scope documentation completeness
- Audit tool access control and privilege management compliance
- Audit log protection, integrity, and retention compliance
- Test data lifecycle management and disposal documentation
- Evidence collection for test data protection and audit programme reviews

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
3. Summary Dashboard - Compliance overview and key metrics
4. Evidence Register - Audit evidence tracking
5. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 2 domains covering Testing and Audit Protection controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a83334_2_audit_activity_management.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a83334_2_audit_activity_management.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a83334_2_audit_activity_management.py --date 20250115

Output:
    File: ISMS-IMP-A.8.33-34.2_Audit_Activity_Management_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.33-34
Assessment Domain:    2 of 2 (Audit Activity Management Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.33-34: Testing and Audit Protection Policy (Governance)
    - ISMS-IMP-A.8.33-34.1: Test Data Protection Assessment (Domain 1)
    - ISMS-IMP-A.8.33-34.2: Audit Activity Management Assessment (Domain 2)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.33-34.2 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive testing and audit protection details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review test data protection procedures and audit activity management annually or when testing methodologies change, new audit tools are deployed, or test data handling incidents are identified.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
import sys
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.33-34.2"
WORKBOOK_NAME = "Audit Activity Management Assessment"
CONTROL_ID = "A.8.33-34"
CONTROL_NAME = "Testing and Audit Protection"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
INFO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Gray for sample rows

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = "\u2705"          # Green check
WARNING = "\u26A0\uFE0F"  # Warning triangle
XMARK = "\u274C"          # Red cross
DASH = "\u2014"           # Em dash


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete all assessment sheets in order, starting with Audit Activity Register.', '2. For each item, evaluate current state against ISO 27001:2022 A.8.33/A.8.34 requirements.', '3. Record all supporting evidence in the Evidence Register sheet.', '4. Use the Summary Dashboard to track overall compliance status.', '5. All user-input cells are highlighted in yellow.', '6. Submit the completed workbook for review and approval via the Approval Sign-Off sheet.', '7. Retain this workbook as part of the ISMS evidence library.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 21

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_audit_activity_register_sheet(ws):
    """Create the Audit Activity Register sheet."""
    ws.title = "Audit Activity Register"

    ws.merge_cells('A1:V1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT ACTIVITY REGISTER")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGNMENT
    ws.row_dimensions[1].height = 35

    ws.cell(row=2, column=1, value="Registry of all planned and completed audit activities")

    headers = [
        "Audit ID", "Audit Name", "Audit Type", "Audit Scope", "Audit Firm Team",
        "Lead Auditor", "Planned Start", "Planned End", "Actual Start", "Actual End",
        "Audit Status", "Management Approval", "Approver", "Approval Date",
        "Systems in Scope", "Data Access Required", "Testing Type", "Findings Count",
        "Critical Findings", "Report Location", "Follow up Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (rows 6-56 = sample + empty rows for user data)
    type_dv = DataValidation(type="list", formula1='"Internal,External-Financial,External-SOC2,Penetration Test,Vulnerability Assessment,Red Team,Regulatory,Compliance"')
    ws.add_data_validation(type_dv)
    type_dv.add('C6:C56')

    status_dv = DataValidation(type="list", formula1='"Planned,In Progress,Completed,Cancelled,On Hold"')
    ws.add_data_validation(status_dv)
    status_dv.add('K6:K56')

    approval_dv = DataValidation(type="list", formula1='"Approved,Pending,Not Required,N/A"')
    ws.add_data_validation(approval_dv)
    approval_dv.add('L6:L56')

    testing_dv = DataValidation(type="list", formula1='"Non-Invasive,Read-Only,Active Testing,Exploitation"')
    ws.add_data_validation(testing_dv)
    testing_dv.add('Q6:Q56')

    followup_dv = DataValidation(type="list", formula1='"Open,In Progress,Closed"')
    ws.add_data_validation(followup_dv)
    followup_dv.add('U6:U56')

    # Row 6: Sample row with example data (GRAY background F2F2F2)
    sample_data = {
        1: "AUD-001",
        2: "Annual Financial Audit 2026",
        3: "External-Financial",
        4: "Financial systems and controls",
        5: "KPMG Audit Team",
        6: "John Smith",
        7: "01.03.2026",
        8: "15.03.2026",
        9: "01.03.2026",
        10: "",
        11: "In Progress",
        12: "Approved",
        13: "CFO",
        14: "15.02.2026",
        15: "SAP, Oracle Financials",
        16: "Financial data (read-only)",
        17: "Read-Only",
        18: "0",
        19: "0",
        20: "/audits/2026/financial",
        21: "Open",
        22: "Scheduled for March 2026"
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=6, column=col, value=value)
        cell.fill = INFO_FILL  # Gray background F2F2F2 for sample row
        cell.border = THIN_BORDER

    # Rows 7-56: Empty rows with yellow background (50 empty FFFFCC rows)
    for row in range(7, 57):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.value = None  # Empty - users choose their own audit IDs

    # Summary Statistics
    ws.cell(row=57, column=1, value="AUDIT STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Audits", "=COUNTA(B7:B56)"),
        ("Planned Audits", '=COUNTIF(K7:K56,"Planned")'),
        ("In Progress", '=COUNTIF(K7:K56,"In Progress")'),
        ("Completed", '=COUNTIF(K7:K56,"Completed")'),
        ("Approved Audits", '=COUNTIF(L7:L56,"Approved")'),
        ("Pending Approval", '=COUNTIF(L7:L56,"Pending")'),
        ("Internal Audits", '=COUNTIF(C7:C56,"Internal")'),
        ("Penetration Tests", '=COUNTIF(C7:C56,"Penetration Test")'),
        ("Total Findings", "=SUM(R7:R56)"),
        ("Total Critical Findings", "=SUM(S7:S56)"),
        ("Open Follow-ups", '=COUNTIF(U7:U56,"Open")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 35, 25, 50, 30, 25, 12, 12, 12, 12, 18, 18, 25, 12, 40, 30, 25, 12, 12, 35, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_tool_authorisation_sheet(ws):
    """Create the Audit Tool Authorisation sheet."""
    ws.title = "Audit Tool Authorisation"

    ws.merge_cells('A1:U1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT TOOL AUTHORISATION")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGNMENT
    ws.row_dimensions[1].height = 35

    ws.cell(row=2, column=1, value="Registry of authorised audit and testing tools")

    headers = [
        "Tool ID", "Tool Name", "Tool Version", "Tool Category", "Vendor Source",
        "Tool Owner", "Authorisation Status", "Authorisation Date", "Authorised By",
        "Risk Level", "Authorised Use Cases", "Restrictions", "Required Approvals",
        "Storage Location", "Access Restricted To", "Last Security Review",
        "Next Review Due", "Usage Logging Required", "License Status",
        "License Expiry", "Evidence Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (rows 6-55 = empty rows for user data)
    category_dv = DataValidation(type="list", formula1='"Vulnerability Scanner,Penetration Tool,Network Analyser,Web App Scanner,Forensic Tool,Credential Tester,Exploitation Framework,Other"')
    ws.add_data_validation(category_dv)
    category_dv.add('D6:D56')

    auth_dv = DataValidation(type="list", formula1='"Authorised,Pending,Unauthorised,Prohibited,N/A"')
    ws.add_data_validation(auth_dv)
    auth_dv.add('G6:G56')

    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('J6:J56')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('R6:R56')

    license_dv = DataValidation(type="list", formula1='"Valid,Expired,N/A"')
    ws.add_data_validation(license_dv)
    license_dv.add('S6:S56')

    # Row 6: Sample row with example data (GRAY background F2F2F2)
    sample_data = {
        1: "TOOL-001",
        2: "Nmap",
        3: "Network Scanner",
        4: "Insecure",
        5: "Host discovery and port scanning",
        6: "Nmap Development Team",
        7: "Authorised",
        8: "Security Team",
        9: "01.01.2026",
        10: "Medium",
        11: "Accidental network disruption",
        12: "Scan only approved test networks",
        13: "Controlled deployment",
        14: "Yes",
        15: "Audit Team Leader",
        16: "Penetration Test",
        17: "PEN-2026-001",
        18: "Yes",
        19: "Open Source",
        20: "N/A"
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=6, column=col, value=value)
        cell.fill = INFO_FILL  # Gray background F2F2F2 for sample row
        cell.border = THIN_BORDER

    # Rows 7-56: Empty rows with yellow background (50 empty FFFFCC rows)
    for row in range(7, 57):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.value = None  # Empty - users choose their own tool IDs

    # Summary Statistics
    ws.cell(row=57, column=1, value="TOOL RISK SUMMARY").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Tools", "=COUNTA(B6:B56)"),
        ("Authorised Tools", '=COUNTIF(G6:G56,"Authorised")'),
        ("Pending Authorisation", '=COUNTIF(G6:G56,"Pending")'),
        ("Unauthorised Tools", '=COUNTIF(G6:G56,"Unauthorised")'),
        ("Prohibited Tools", '=COUNTIF(G6:G56,"Prohibited")'),
        ("High Risk Tools", '=COUNTIF(J6:J56,"High")'),
        ("Medium Risk Tools", '=COUNTIF(J6:J56,"Medium")'),
        ("Low Risk Tools", '=COUNTIF(J6:J56,"Low")'),
        ("Expired Licenses", '=COUNTIF(S6:S56,"Expired")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 30, 15, 25, 25, 25, 18, 12, 25, 15, 50, 40, 30, 35, 30, 12, 12, 12, 18, 12, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_access_tracking_sheet(ws):
    """Create the Audit Access Tracking sheet."""
    ws.title = "Audit Access Tracking"

    ws.merge_cells('A1:X1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT ACCESS TRACKING")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGNMENT
    ws.row_dimensions[1].height = 35

    ws.cell(row=2, column=1, value="Registry of auditor access to systems and data")

    headers = [
        "Access ID", "Auditor Name", "Auditor Organisation", "Associated Audit",
        "Access Type", "Systems Accessed", "Data Classification Accessed",
        "Access Requested Date", "Access Start Date", "Access End Date",
        "Actual Revocation Date", "Access Status", "Approval Status", "Approver",
        "Approval Date", "Access Logging Enabled", "NDA Signed", "NDA Reference",
        "Supervision Required", "Supervisor", "Multi Factor Auth Required",
        "Revocation Confirmation", "Notes", "Device Security Verified"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (rows 6-55 = empty rows for user data)
    access_type_dv = DataValidation(type="list", formula1='"Read-Only,Read-Write,Admin,Physical,Remote-VPN"')
    ws.add_data_validation(access_type_dv)
    access_type_dv.add('E6:E56')

    classification_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted,N/A"')
    ws.add_data_validation(classification_dv)
    classification_dv.add('G6:G56')

    status_dv = DataValidation(type="list", formula1='"Active,Revoked,Expired"')
    ws.add_data_validation(status_dv)
    status_dv.add('L6:L56')

    approval_dv = DataValidation(type="list", formula1='"Approved,Pending,Denied,N/A"')
    ws.add_data_validation(approval_dv)
    approval_dv.add('M6:M56')

    ynp_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(ynp_dv)
    ynp_dv.add('P6:P56')

    yn_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('Q6:Q56')
    yn_dv.add('S6:S56')
    yn_dv.add('U6:U56')

    device_sec_dv = DataValidation(type="list", formula1='"Yes,Pending,N/A"')
    ws.add_data_validation(device_sec_dv)
    device_sec_dv.add('X6:X56')

    # Row 6: Sample row with example data (GRAY background F2F2F2)
    sample_data = {
        1: "ACC-001",
        2: "Jane Auditor",
        3: "External Audit Firm",
        4: "AUD-001",
        5: "Read-Only",
        6: "Financial Systems",
        7: "Confidential",
        8: "01.03.2026",
        9: "01.03.2026",
        10: "15.03.2026",
        11: "",
        12: "Active",
        13: "Approved",
        14: "IT Security Manager",
        15: "28.02.2026",
        16: "Yes",
        17: "Yes",
        18: "NDA-2026-001",
        19: "No",
        20: "",
        21: "Yes",
        22: "",
        23: "Access granted for financial audit",
        24: "Yes — MDM enrolled, patched, AV active"
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=6, column=col, value=value)
        cell.fill = INFO_FILL  # Gray background F2F2F2 for sample row
        cell.border = THIN_BORDER

    # Rows 7-56: Empty rows with yellow background (50 empty FFFFCC rows)
    for row in range(7, 57):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.value = None  # Empty - users choose their own access IDs

    # Summary Statistics
    ws.cell(row=57, column=1, value="ACCESS STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Access Grants", "=COUNTA(B6:B56)"),
        ("Active Access", '=COUNTIF(L6:L56,"Active")'),
        ("Revoked Access", '=COUNTIF(L6:L56,"Revoked")'),
        ("Expired Access", '=COUNTIF(L6:L56,"Expired")'),
        ("Approved Access", '=COUNTIF(M6:M56,"Approved")'),
        ("Pending Approval", '=COUNTIF(M6:M56,"Pending")'),
        ("Read-Only Access", '=COUNTIF(E6:E56,"Read-Only")'),
        ("Admin Access", '=COUNTIF(E6:E56,"Admin")'),
        ("Restricted Data Access", '=COUNTIF(G6:G56,"Restricted")'),
        ("NDA Coverage (Yes)", '=COUNTIF(Q6:Q56,"Yes")'),
        ("Access Logging (Yes)", '=COUNTIF(P6:P56,"Yes")'),
        ("MFA Required (Yes)", '=COUNTIF(U6:U56,"Yes")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 25, 25, 20, 20, 40, 18, 12, 12, 12, 12, 18, 18, 25, 12, 12, 12, 25, 12, 25, 12, 20, 40, 25]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_disruption_mitigation_sheet(ws):
    """Create the Disruption Mitigation Plans sheet."""
    ws.title = "Disruption Mitigation Plans"

    ws.merge_cells('A1:W1')
    title_cell = ws.cell(row=1, column=1, value="DISRUPTION MITIGATION PLANS")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGNMENT
    ws.row_dimensions[1].height = 35

    ws.cell(row=2, column=1, value="Protection measures for production systems during audit testing")

    headers = [
        "System ID", "System Name", "System Criticality", "Business Owner",
        "Technical Owner", "Associated Audits", "Primary Mitigation Strategy",
        "Testing Restrictions", "Permitted Testing Window", "Maximum Test Duration",
        "Backup Required Before Test", "Recovery Point Objective", "Recovery Time Objective",
        "Rollback Procedure Location", "Rollback Last Tested", "Escalation Contact",
        "Escalation Phone", "Secondary Contact", "Monitoring Enhancement",
        "Incident Response Plan", "Last Mitigation Review", "Review Due Date", "Evidence Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (rows 6-55 = empty rows for user data)
    criticality_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low,N/A"')
    ws.add_data_validation(criticality_dv)
    criticality_dv.add('C6:C56')

    strategy_dv = DataValidation(type="list", formula1='"Staging First,Off-Hours,Rate Limiting,Scope Exclusion,Enhanced Monitoring,Standby Recovery,Read-Only Only,Multi-Strategy"')
    ws.add_data_validation(strategy_dv)
    strategy_dv.add('G6:G56')

    backup_dv = DataValidation(type="list", formula1='"Yes,No,Already Scheduled,N/A"')
    ws.add_data_validation(backup_dv)
    backup_dv.add('K6:K56')

    # Row 6: Sample row with example data (GRAY background F2F2F2)
    sample_data = {
        1: "SYS-001",
        2: "Financial System (SAP)",
        3: "Critical",
        4: "CFO",
        5: "SAP Team Lead",
        6: "AUD-001",
        7: "Staging First",
        8: "No destructive testing, read-only queries only",
        9: "Saturday 20:00-Sunday 06:00",
        10: "2 hours",
        11: "Yes",
        12: "1 hour",
        13: "15 minutes",
        14: "/procedures/sap-rollback.pdf",
        15: "01.02.2026",
        16: "IT Director",
        17: "+41 XX XXX XX XX",
        18: "SAP Manager",
        19: "Enhanced real-time monitoring during test",
        20: "INC-PLAN-001",
        21: "01.01.2026",
        22: "01.07.2026",
        23: "EV-001"
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=6, column=col, value=value)
        cell.fill = INFO_FILL  # Gray background F2F2F2 for sample row
        cell.border = THIN_BORDER

    # Rows 7-56: Empty rows with yellow background (50 empty FFFFCC rows)
    for row in range(7, 57):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.value = None  # Empty - users choose their own system IDs

    # Summary Statistics
    ws.cell(row=57, column=1, value="MITIGATION STRATEGY SUMMARY").font = Font(bold=True, size=12)

    strategies = ["Staging First", "Off-Hours", "Rate Limiting", "Scope Exclusion", "Enhanced Monitoring", "Standby Recovery", "Read-Only Only", "Multi-Strategy"]
    row = 59
    for strategy in strategies:
        ws.cell(row=row, column=1, value=strategy).border = THIN_BORDER
        ws.cell(row=row, column=2, value=f'=COUNTIF(G6:G56,"{strategy}")').border = THIN_BORDER
        row += 1

    ws.cell(row=row + 1, column=1, value="SYSTEM COVERAGE STATISTICS").font = Font(bold=True, size=12)
    row += 3

    coverage_items = [
        ("Total Systems", "=COUNTA(B6:B56)"),
        ("Critical Systems", '=COUNTIF(C6:C56,"Critical")'),
        ("High Criticality", '=COUNTIF(C6:C56,"High")'),
        ("Medium Criticality", '=COUNTIF(C6:C56,"Medium")'),
        ("Low Criticality", '=COUNTIF(C6:C56,"Low")'),
        ("With Rollback Procedure", '=COUNTA(N6:N56)'),
        ("Backup Required", '=COUNTIF(K6:K56,"Yes")'),
    ]

    for label, formula in coverage_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 30, 15, 25, 25, 30, 35, 50, 25, 15, 12, 15, 15, 35, 12, 25, 20, 25, 30, 35, 12, 12, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_evidence_protection_sheet(ws):
    """Create the Audit Evidence Protection sheet."""
    ws.title = "Audit Evidence Protection"

    ws.merge_cells('A1:W1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT EVIDENCE PROTECTION")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGNMENT
    ws.row_dimensions[1].height = 35

    ws.cell(row=2, column=1, value="Secure handling and retention of audit evidence")

    headers = [
        "Evidence Category ID", "Evidence Category", "Evidence Description",
        "Sensitivity Classification", "Example Documents", "Primary Storage Location",
        "Backup Location", "Encryption at Rest", "Encryption in Transit",
        "Access Control Type", "Authorised Accessors", "Retention Period",
        "Retention Start Event", "Destruction Method", "Destruction Approval",
        "Chain of Custody Required", "Chain of Custody Process", "Legal Hold Applicable",
        "Legal Hold Contact", "Integrity Verification", "Last Access Review",
        "Evidence Owner", "Evidence Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (rows 5-64: 10 predefined + 50 empty)
    classification_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted,N/A"')
    ws.add_data_validation(classification_dv)
    classification_dv.add('D5:D64')

    enc_rest_dv = DataValidation(type="list", formula1='"AES-256,AES-128,BitLocker,None,Other"')
    ws.add_data_validation(enc_rest_dv)
    enc_rest_dv.add('H5:H64')

    enc_transit_dv = DataValidation(type="list", formula1='"TLS 1.3,TLS 1.2,VPN,None,Other"')
    ws.add_data_validation(enc_transit_dv)
    enc_transit_dv.add('I5:I64')

    access_dv = DataValidation(type="list", formula1='"RBAC,ACL,Individual Permissions,None"')
    ws.add_data_validation(access_dv)
    access_dv.add('J5:J64')

    retention_dv = DataValidation(type="list", formula1='"1 Year,2 Years,3 Years,5 Years,7 Years,Permanent,Legal Hold"')
    ws.add_data_validation(retention_dv)
    retention_dv.add('L5:L54')

    destruction_dv = DataValidation(type="list", formula1='"Secure Delete,Shredding,Degaussing,Certified Destruction"')
    ws.add_data_validation(destruction_dv)
    destruction_dv.add('N5:N64')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('P5:P64')

    legal_dv = DataValidation(type="list", formula1='"Yes,No,Potentially"')
    ws.add_data_validation(legal_dv)
    legal_dv.add('R5:R64')

    integrity_dv = DataValidation(type="list", formula1='"Hashing,Digital Signatures,Checksums,None"')
    ws.add_data_validation(integrity_dv)
    integrity_dv.add('T5:T64')

    # Sample row (row 5) with complete data - GRAY background
    INFO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    sample_data = {
        1: "EVCAT-001",
        2: "Penetration Test Reports",
        3: "Technical findings from pen tests",
        4: "Restricted",
        5: "Pen test report Q1 2026",
        6: "/secure/audit/pentest/",
        7: "/backup/audit/pentest/",
        8: "AES-256",
        9: "TLS 1.3",
        10: "RBAC",
        11: "Security Team, CISO",
        12: "3 Years",
        13: "Test completion",
        14: "Secure Delete",
        15: "CISO",
        16: "Yes",
        17: "Documented chain process",
        18: "No",
        19: "",
        20: "Hashing",
        21: "15.01.2026",
        22: "Security Manager",
        23: "SEC-AUDIT-001",
    }

    for col, value in sample_data.items():
        cell = ws.cell(row=6, column=col, value=value)
        cell.fill = INFO_FILL  # Gray background F2F2F2 for sample row
        cell.border = THIN_BORDER

    # Empty rows (50 rows for user data: rows 7-56) - YELLOW background
    for row in range(7, 57):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.value = None  # Empty - users choose their own category IDs

    # Summary Statistics
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=67, column=1, value="PROTECTION STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Evidence Categories", "=COUNTA(B5:B54)"),
        ("Restricted Categories", '=COUNTIF(D5:D64,"Restricted")'),
        ("Confidential Categories", '=COUNTIF(D5:D64,"Confidential")'),
        ("Encryption at Rest (AES-256)", '=COUNTIF(H5:H64,"AES-256")'),
        ("Encryption in Transit (TLS 1.3)", '=COUNTIF(I5:I64,"TLS 1.3")'),
        ("Chain of Custody Required", '=COUNTIF(P5:P64,"Yes")'),
        ("Legal Hold Applicable", '=COUNTIF(R5:R64,"Yes")'),
    ]

    row = 69
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 30, 50, 18, 40, 35, 35, 18, 18, 25, 40, 18, 25, 25, 25, 12, 40, 12, 25, 25, 12, 25, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet."""
    ws.title = "Summary Dashboard"
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)


    # Header (Row 1)
    ws.merge_cells("A1:G1")
    ws["A1"] = "AUDIT ACTIVITY MANAGEMENT \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: CONTROL_REF
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 — Control A.8.33-34: Testing Information Protection and Audit Logging"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty (for standardisation)

    # TABLE 1: Compliance Summary (Row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

    # Column Headers (Row 5)
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows — one per assessment sheet
    # Each row references the actual status column DV in that sheet
    area_configs = [
        # (Area Name, Status Column, Status Values: [Good, Partial, Bad])
        ("Audit Activity Register", "L", ["Approved", "Pending", "Not Required"]),
        ("Audit Tool Authorisation", "G", ["Authorised", "Pending", "Unauthorised+COUNTIF('Audit Tool Authorisation'!G7:G56,\"Prohibited\")"]),
        ("Audit Access Tracking", "M", ["Approved", "Pending", "Denied"]),
        ("Disruption Mitigation Plans", "K", ["Yes", "Already Scheduled", "No"]),  # Backup_Required_Before_Test - compliance-oriented column
        ("Audit Evidence Protection", "D", ["Restricted", "Confidential", "Internal+COUNTIF('Audit Evidence Protection'!D7:D56,\"Public\")"]),  # Sensitivity - not ideal but best available
    ]

    for i, (area_name, status_col, status_values) in enumerate(area_configs):
        row = 6 + i

        # Column A: Area name
        ws.cell(row=row, column=1, value=area_name).border = THIN_BORDER

        # Column B: Total Items (COUNT formula)
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTA('{area_name}'!B7:B56)"
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = Font(color="000000")

        # Column C: Good status (Approved, Authorised, Critical, Restricted)
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f'=COUNTIF(\'{area_name}\'!{status_col}7:{status_col}56,"{status_values[0]}")'
        cell_c.border = THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = Font(color="000000")

        # Column D: Partial status (Pending, High, Confidential)
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f'=COUNTIF(\'{area_name}\'!{status_col}7:{status_col}56,"{status_values[1]}")'
        cell_d.border = THIN_BORDER
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = Font(color="000000")

        # Column E: Bad status (Not Required, Denied, Unauthorised+Prohibited, Medium+Low, Internal+Public)
        # Some need multiple COUNTIF for multiple bad values
        cell_e = ws.cell(row=row, column=5)
        bad_value = status_values[2]

        # Parse compound values (e.g., "Unauthorised+COUNTIF(...)")
        if '+COUNTIF' in bad_value:
            # Extract the base value and the additional formula
            parts = bad_value.split('+COUNTIF')
            base_val = parts[0]
            additional_formula = 'COUNTIF' + parts[1]
            cell_e.value = f'=COUNTIF(\'{area_name}\'!{status_col}7:{status_col}56,"{base_val}")+{additional_formula}'
        else:
            cell_e.value = f'=COUNTIF(\'{area_name}\'!{status_col}7:{status_col}56,"{bad_value}")'

        cell_e.border = THIN_BORDER
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = Font(color="000000")

        # Column F: N/A
        ws.cell(row=row, column=6, value=f'=COUNTIF(\'{area_name}\'!{status_col}7:{status_col}56,"N/A")').border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6).font = Font(color="000000")

        # Column G: Compliance %
        cell = ws.cell(row=row, column=7)
        cell.value = f'=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))'
        cell.number_format = "0.0%"
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(color="000000")

    # TOTAL row
    total_row = 6 + len(area_configs)  # Row 11 (5 area_configs)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    # Compliance % formula
    cell = ws.cell(row=total_row, column=7)
    cell.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    cell.number_format = "0.0%"
    cell.font = Font(bold=True, color="000000")
    cell.fill = grey_fill
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")

    # TABLE 2: Key Metrics (exploiting ALL DVs across all sheets)
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{metrics_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=metrics_start, column=c).border = THIN_BORDER

    metric_headers = ["Metric", "Value", "", "", "", "", ""]
    for col, header in enumerate(metric_headers, 1):
        cell = ws.cell(row=metrics_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # All metrics with formulas exploiting DVs from all 5 sheets
    metrics = [
        # Audit Activity Metrics
        ("Total Audit Activities", "=COUNTA('Audit Activity Register'!B7:B56)"),
        ("Pending Approval", '=COUNTIF(\'Audit Activity Register\'!L7:L56,"Pending")'),
        ("Vulnerability Scans", '=COUNTIF(\'Audit Activity Register\'!C7:C56,"Vulnerability Assessment")'),
        ("Penetration Tests", '=COUNTIF(\'Audit Activity Register\'!C7:C56,"Penetration Test")'),
        # Tool Authorisation Metrics
        ("Total Audit Tools", "=COUNTA('Audit Tool Authorisation'!B7:B56)"),
        ("Authorised Tools", '=COUNTIF(\'Audit Tool Authorisation\'!G7:G56,"Authorised")'),
        ("Unauthorised Tools", '=COUNTIF(\'Audit Tool Authorisation\'!G7:G56,"Unauthorised")'),
        # Access Tracking Metrics
        ("Total Access Requests", "=COUNTA('Audit Access Tracking'!B7:B56)"),
        ("Approved Access", '=COUNTIF(\'Audit Access Tracking\'!M7:M56,"Approved")'),
        ("Denied Access", '=COUNTIF(\'Audit Access Tracking\'!M7:M56,"Denied")'),
        # Disruption Metrics
        ("Critical Systems", '=COUNTIF(\'Disruption Mitigation Plans\'!C7:C56,"Critical")'),
        ("Plans Requiring Backup", '=COUNTIF(\'Disruption Mitigation Plans\'!K7:K56,"Yes")'),
        # Evidence Protection Metrics
        ("Total Evidence Items", "=COUNTA('Audit Evidence Protection'!B7:B56)"),
        ("Restricted Evidence", '=COUNTIF(\'Audit Evidence Protection\'!D7:D56,"Restricted")'),
        ("Legal Hold Applicable", '=COUNTIF(\'Audit Evidence Protection\'!R7:R56,"Yes")'),
    ]

    row = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = THIN_BORDER
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TABLE 2: 2 empty buffer rows (for manual additions)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TABLE 3: Critical Findings
    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=crit_start, column=c).border = THIN_BORDER

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Critical findings based on control A.8.34 requirements (audit testing protection)
    findings = [
        ("Audit Tool Authorisation", "Tools without authorisation", '=COUNTIF(\'Audit Tool Authorisation\'!G7:G56,"Unauthorised")', "Critical", "Immediate"),
        ("Audit Tool Authorisation", "Prohibited tools in use", '=COUNTIF(\'Audit Tool Authorisation\'!G7:G56,"Prohibited")', "Critical", "Immediate"),
        ("Disruption Mitigation Plans", "Critical systems without plans", '=COUNTIF(\'Disruption Mitigation Plans\'!C7:C56,"Critical")', "Critical", "Immediate"),
        ("Audit Tool Authorisation", "High-risk tools", '=COUNTIF(\'Audit Tool Authorisation\'!J7:J56,"High")', "High", "Urgent"),
        ("Audit Access Tracking", "Access denied (approval gap)", '=COUNTIF(\'Audit Access Tracking\'!M7:M56,"Denied")', "High", "Urgent"),
        ("Disruption Mitigation Plans", "High-risk systems without plans", '=COUNTIF(\'Disruption Mitigation Plans\'!C7:C56,"High")', "High", "Urgent"),
        ("Audit Tool Authorisation", "Usage logging not enabled", '=COUNTIF(\'Audit Tool Authorisation\'!R7:R56,"No")', "Medium", "Plan"),
        ("Audit Tool Authorisation", "Expired tool licenses", '=COUNTIF(\'Audit Tool Authorisation\'!S7:S56,"Expired")', "Medium", "Plan"),
        ("Audit Access Tracking", "Expired access not revoked", '=COUNTIF(\'Audit Access Tracking\'!L7:L56,"Expired")', "Medium", "Plan"),
        ("Disruption Mitigation Plans", "No backup before testing", '=COUNTIF(\'Disruption Mitigation Plans\'!K7:K56,"No")', "Medium", "Plan"),
    ]

    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # TABLE 3: 2 empty buffer rows (for manual additions)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create the Evidence Register sheet — standard 8-column format."""
    ws.title = "Evidence Register"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1)
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Apply borders to all cells in title merged range (A1:E1)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column Headers (Row 4)
    headers = [
        ("Evidence ID", 15),
        ("Assessment Area", 25),
        ("Evidence Type", 22),
        ("Description", 40),
        ("Location/Path", 45),
        ("Date Collected", 16),
        ("Collected By", 20),
        ("Verification Status", 22),
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")  # White text on dark blue
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    type_dv.add("C5:C105")

    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("H5:H105")

    # Sample row (row 5) with example data
    # Sample row (row 5) with complete data and GRAY background
    sample_data_full = {
        1: "EV-001",
        2: "Audit Activity Register",
        3: "Audit log",
        4: "Completed audit activities for 2026",
        5: "/evidence/audits/2026/register.xlsx",
        6: "01.03.2026",
        7: "Audit Coordinator",
        8: "Verified"
    }
    for col, value in sample_data_full.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = INFO_FILL  # Grey background (F2F2F2) for sample row - visual distinction
        cell.border = border

    # Empty data rows (rows 6-105, 100 rows) - Exception: Evidence Register has 100 rows
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = INPUT_FILL
            cell.border = border
            cell.value = None  # Empty - users choose their own evidence IDs

    # Freeze
    ws.freeze_panes = "A5"


def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — standard pattern."""
    ws.title = "Approval Sign-Off"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1)
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Apply borders to all cells in title merged range (A1:E1)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # ASSESSMENT SUMMARY banner (Row 3)
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to all cells in banner merged range (A3:E3)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G11"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        # Apply borders to all cells in merged range
        for c in range(2, 6):  # Columns B-E
            ws.cell(row=row, column=c).border = border
        row += 1

    # Status dropdown on Assessment Status
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"B{row - 1}")

    # 3 Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Apply borders to all cells in approver banner merged range (A:E)
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            # Apply borders to all cells in merged range
            for c in range(2, 6):  # Columns B-E
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    # Apply borders to all cells in merged range
    for c in range(2, 6):  # Columns B-E
        ws.cell(row=row, column=c).border = border

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Apply borders to all cells in NEXT REVIEW banner merged range (A:E)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        # Apply borders to all cells in merged range
        for c in range(2, 6):  # Columns B-E
            ws.cell(row=row, column=c).border = border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    default_sheet = wb.active
    default_sheet.sheet_view.showGridLines = False

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_audit_activity_register_sheet(wb.create_sheet())
    create_audit_tool_authorisation_sheet(wb.create_sheet())
    create_audit_access_tracking_sheet(wb.create_sheet())
    create_disruption_mitigation_sheet(wb.create_sheet())
    create_audit_evidence_protection_sheet(wb.create_sheet())
    create_evidence_register(wb.create_sheet())
    create_summary_dashboard_sheet(wb.create_sheet())
    create_approval_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    # Save workbook
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")



# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
