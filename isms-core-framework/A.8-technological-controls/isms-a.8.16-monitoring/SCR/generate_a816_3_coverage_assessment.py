#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.16.3 - Coverage Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities
Assessment Domain 3 of 5: Monitoring Coverage Across Assets, Networks, and Applications

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific asset inventory, network topology, and
coverage requirements.

Key customisation areas:
1. Asset tiers and criticality definitions (match your classification)
2. Network segment inventory (specific to your topology)
3. Cloud environment scope (AWS, Azure, GCP per your infrastructure)
4. Coverage thresholds per tier (aligned with your risk appetite)
5. Blind spot identification criteria (based on your threat model)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for monitoring)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
monitoring coverage across all critical assets, network segments, user activities,
applications, and cloud environments against ISO 27001:2022 Control A.8.16 requirements.

**Purpose:**
Enables systematic assessment of monitoring coverage completeness, identification
of blind spots, and validation that all critical systems have appropriate monitoring
per their criticality tier (Tier 1: 100%, Tier 2: 90%, Tier 3: 70%).

**Assessment Scope:**
- Asset coverage (endpoints, servers, infrastructure)
- Network segment coverage (internal, DMZ, remote access)
- User activity monitoring (privileged users, normal users)
- Application coverage (business-critical applications)
- Cloud environment monitoring (IaaS, PaaS, SaaS)
- Gap analysis and blind spot identification
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and coverage standards
2. Asset Coverage - Endpoint, server, and infrastructure monitoring
3. Network Coverage - Network segment and traffic monitoring
4. User Activity Coverage - User behavior and privileged access monitoring
5. Application Coverage - Application-level monitoring and logging
6. Cloud Coverage - Cloud environment monitoring across all providers
7. Summary Dashboard - Overall coverage metrics and gap summary
8. Evidence Register - Audit evidence tracking and documentation
9. Approval Sign-Off - Multi-level approval workflow

**Key Features:**
- Data validation with criticality tier dropdown lists
- Conditional formatting for coverage status visualization
- Automated coverage percentage calculations by tier
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with A.8.16 Compliance Dashboard

**Integration:**
This assessment feeds into ISMS-IMP-A.8.16.5 Compliance Dashboard, which
consolidates data from all five monitoring assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a816_3_coverage_assessment.py

Output:
    File: ISMS_A_8_16_3_Coverage_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Inventory ALL assets (endpoints, servers, infrastructure)
    2. Document criticality tier for each asset (Tier 1/2/3)
    3. Assess current monitoring coverage per asset
    4. Identify coverage gaps and blind spots
    5. Validate network segment monitoring completeness
    6. Review user activity monitoring coverage
    7. Assess application-level monitoring
    8. Evaluate cloud environment monitoring
    9. Calculate coverage percentages by tier
    10. Define remediation actions for coverage gaps
    11. Collect and link audit evidence
    12. Obtain stakeholder approvals
    13. Feed results into A.8.16.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.16
Assessment Domain:    3 of 5 (Coverage Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.16: Monitoring Activities Policy (Governance)
    - ISMS-POL-A.8.16-S2.1: Monitoring Infrastructure Requirements (Coverage)
    - ISMS-IMP-A.8.16.3: Coverage Assessment Implementation Guide
    - ISMS-IMP-A.8.16.1: Monitoring Infrastructure Assessment (Domain 1)
    - ISMS-IMP-A.8.16.2: Baseline & Detection Assessment (Domain 2)
    - ISMS-IMP-A.8.16.4: Alert Management Assessment (Domain 4)
    - ISMS-IMP-A.8.16.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.16.3 specification
    - Supports comprehensive coverage evaluation across all asset types
    - Integrated with A.8.16.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Coverage Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This is not cargo cult ISMS. Coverage gaps are security blind spots. If a
critical system is not monitored, threats affecting it are invisible. Complete
coverage assessment requires INVENTORY of everything that processes, stores,
or transmits data.

**Criticality Tier Requirements (ISMS-POL-A.8.16-S2.1):**
- Tier 1 (Critical): 100% monitoring coverage REQUIRED
- Tier 2 (High): 90% monitoring coverage target
- Tier 3 (Standard): 70% monitoring coverage target

Tier 1 systems with <100% coverage represent CRITICAL GAPS requiring immediate
remediation. No exceptions.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will expect:
- Complete asset inventory with criticality classifications
- Evidence of monitoring deployment per asset
- Coverage gap analysis with remediation plans
- Coverage metrics by criticality tier

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Complete asset inventory (endpoints, servers, applications)
- Network topology and segmentation
- Coverage gaps and monitoring blind spots
- Cloud architecture and service inventory

Handle in accordance with your organisation's data classification policies.

**Common Coverage Gaps (from field experience):**
- Legacy systems assumed "air-gapped" but actually networked
- Contractor/vendor endpoints not in asset inventory
- Development/test environments (often attacked first)
- OT/ICS systems (different monitoring requirements)
- Shadow IT (SaaS applications not centrally managed)
- Mobile devices (BYOD and corporate)
- IoT devices (printers, cameras, sensors)

Every organisation has blind spots. This assessment finds them.

**Maintenance:**
Review and update assessment:
- Monthly: Add new assets as they're deployed
- Quarterly: Validate coverage for all Tier 1 systems
- Semi-annually: Complete coverage assessment for all tiers
- Annually: Full reassessment with updated criticality classifications
- Ad-hoc: When new systems deployed or infrastructure changes

**Quality Assurance:**
Have IT asset management and SOC teams validate assessments together. Asset
inventory accuracy is critical - if it's not in the inventory, it's not monitored.

**Integration Note:**
Coverage assessment should be synchronized with:
- IT Asset Management (ITAM) inventory
- Configuration Management Database (CMDB)
- Business Impact Analysis (BIA) - for criticality tiers
- Cloud resource inventory (auto-discovery where possible)

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    import openpyxl
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.16.3"
WORKBOOK_NAME = "Coverage Assessment"
CONTROL_ID = "A.8.16"
CONTROL_NAME = "Monitoring Activities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Output directory (WKBK/ sibling of SCR/)
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)



def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.16.3 specification
    sheets = [
        "Instructions & Legend",
        "1. Asset Coverage",
        "2. Network Coverage",
        "3. User Identity Coverage",
        "4. Application Coverage",
        "5. Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "example_row": {
            "font": Font(name="Calibri", size=9, italic=True, color="808080"),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None,
            italic=style_dict["font"].italic if hasattr(style_dict["font"], 'italic') else False
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial"',
            allow_blank=False
        ),
        'yes_partial_no': DataValidation(
            type="list",
            formula1='"Yes,Partial,No"',
            allow_blank=False
        ),
        'asset_type': DataValidation(
            type="list",
            formula1='"Server,Network Device,Security Device,Endpoint,Cloud Resource,Database,Application,Container,IoT Device,Other"',
            allow_blank=False
        ),
        'data_classification': DataValidation(
            type="list",
            formula1='"Confidential,Internal,Public"',
            allow_blank=False
        ),
        'criticality': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'regulatory_scope': DataValidation(
            type="list",
            formula1='"PCI-DSS,HIPAA,GDPR,SOX,Multiple,None"',
            allow_blank=False
        ),
        'monitoring_required': DataValidation(
            type="list",
            formula1='"Mandatory,Recommended,Optional,N/A"',
            allow_blank=False
        ),
        'coverage_status': DataValidation(
            type="list",
            formula1='"\u2705 Full Coverage,\u26A0\uFE0F Partial Coverage,\u274C No Coverage,N/A"',
            allow_blank=False
        ),
        'segment_type': DataValidation(
            type="list",
            formula1='"Production,DMZ,Internal,Management,Guest,Partner,Development,Test,Cloud VPC,Other"',
            allow_blank=False
        ),
        'perimeter_monitoring': DataValidation(
            type="list",
            formula1='"Firewall,IDS/IPS,Both,None"',
            allow_blank=False
        ),
        'isolation_status': DataValidation(
            type="list",
            formula1='"Isolated,Semi-Isolated,Open"',
            allow_blank=False
        ),
        'identity_system_type': DataValidation(
            type="list",
            formula1='"Active Directory,Microsoft Entra ID (formerly Azure AD),LDAP,SAML IdP,OAuth Provider,Database Auth,Application-Specific,Other"',
            allow_blank=False
        ),
        'application_type': DataValidation(
            type="list",
            formula1='"Web Application,Database,API,Microservice,SaaS,Mobile App,Desktop App,Other"',
            allow_blank=False
        ),
        'gap_category': DataValidation(
            type="list",
            formula1='"Asset Not Monitored,Log Source Missing,Network Segment Gap,User/Identity Gap,Application Gap,Detection Gap,Other"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Deferred,Accepted"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1='"\u2705 Compliant,\u26A0\uFE0F Partial,\u274C Non-Compliant,N/A"',
            allow_blank=False
        ),
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
    }
    
    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete comprehensive asset inventory (Sheet 1).', '2. Document all network segments and coverage (Sheet 2).', '3. Assess user and identity system monitoring (Sheet 3).', '4. Inventory application monitoring coverage (Sheet 4).', '5. Identify and document all coverage gaps (Sheet 5).', '6. Review Summary Dashboard for overall coverage status.', '7. Prioritise gaps by criticality and regulatory requirements.', '8. Document evidence in Evidence Register.', '9. Obtain approvals via Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_asset_coverage_sheet(ws, styles):
    """Create Asset Inventory & Coverage assessment sheet."""
    ws.row_dimensions[1].height = 35
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:W1")
    ws["A1"] = "1. ASSET INVENTORY & MONITORING COVERAGE"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:W2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.1 - Assess asset coverage"
    apply_style(ws["A2"], styles["subheader"])

    ws["A3"] = "Are all organisational assets inventoried and appropriate assets monitored?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:W3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Asset ID", 15),
        ("B", "Asset Name", 28),
        ("C", "Asset Type", 22),
        ("D", "Operating System", 22),
        ("E", "Location", 18),
        ("F", "Business Unit", 20),
        ("G", "Asset Owner", 20),
        ("H", "Data Classification", 18),
        ("I", "Criticality", 15),
        ("J", "Regulatory Scope", 22),
        ("K", "Monitoring Required", 16),
        ("L", "Currently Monitored", 16),
        ("M", "Log Types Collected", 30),
        ("N", "Monitoring Platform", 22),
        ("O", "Baseline Established", 16),
        ("P", "Detection Rules Active", 18),
        ("Q", "Last Log Verified", 14),
        ("R", "Coverage Status", 18),
        ("S", "Gap Reason", 30),
        ("T", "Exception Approved", 16),
        ("U", "Target Coverage Date", 14),
        ("V", "Responsible Party", 20),
        ("W", "Notes", 25),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "SRV-001", "DC01-Primary", "Server", "Windows Server 2022", "DataCenter-1",
        "IT Operations", "J. Smith", "Confidential", "Critical", "Multiple",
        "Mandatory", "Yes", "Security, System, Application", "Splunk", "Yes",
        "15", "05.01.2025", "\u2705 Full Coverage", "None", "N/A", "N/A", "SOC Team", "Production DC"
    ]

    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    for data_row in range(8, 58):
        for col_idx in range(1, 24):  # A to W
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

            col_letter = get_column_letter(col_idx)
            if col_letter == 'C':  # Asset Type
                validations['asset_type'].add(cell)
            elif col_letter == 'H':  # Data Classification
                validations['data_classification'].add(cell)
            elif col_letter == 'I':  # Criticality
                validations['criticality'].add(cell)
            elif col_letter == 'J':  # Regulatory Scope
                validations['regulatory_scope'].add(cell)
            elif col_letter == 'K':  # Monitoring Required
                validations['monitoring_required'].add(cell)
            elif col_letter == 'L':  # Currently Monitored
                validations['yes_no_partial'].add(cell)
            elif col_letter == 'O':  # Baseline Established
                validations['yes_no_na'].add(cell)
            elif col_letter == 'R':  # Coverage Status
                validations['coverage_status'].add(cell)
            elif col_letter == 'T':  # Exception Approved
                validations['yes_no_na'].add(cell)

    # Compliance Checklist (Starting Row 59)
    row = 59
    ws.merge_cells(f"A{row}:W{row}")
    ws[f"A{row}"] = "ASSET COVERAGE COMPLIANCE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Complete asset inventory maintained",
        "Asset inventory updated at least quarterly",
        "All assets classified by criticality",
        "All Critical assets monitored (100%)",
        ">80% of High priority assets monitored",
        ">60% of Medium priority assets monitored",
        "All PCI-DSS scope systems monitored",
        "All HIPAA scope systems monitored",
        "All systems handling confidential data monitored",
        "Monitoring coverage gaps documented",
        "Exceptions formally approved",
        "Gap remediation plans exist",
        "Coverage status reported monthly",
        "Asset decommissioning process includes monitoring removal",
        "New assets onboarded to monitoring within 30 days",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:V{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:V{row}")

        ws[f"W{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"W{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"W{row}"])
        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(W61:W75,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Finalize all validations
    for _dv in validations.values():
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: SHEET 3 - NETWORK SEGMENT COVERAGE
# ============================================================================

def create_network_coverage_sheet(ws, styles):
    """Create Network Segment Coverage assessment sheet."""
    ws.row_dimensions[1].height = 35
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:T1")
    ws["A1"] = "2. NETWORK SEGMENT COVERAGE ASSESSMENT"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:T2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.1 - Assess network coverage"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        ("A", "Network Segment/Zone", 28),
        ("B", "Segment Type", 22),
        ("C", "IP Range/CIDR", 20),
        ("D", "VLAN ID", 12),
        ("E", "Number of Assets", 15),
        ("F", "Criticality", 15),
        ("G", "Data Classification", 18),
        ("H", "Perimeter Monitoring", 18),
        ("I", "Flow Monitoring", 16),
        ("J", "DNS Monitoring", 16),
        ("K", "Endpoint Monitoring", 18),
        ("L", "Log Collection Active", 18),
        ("M", "Network Tap/SPAN", 16),
        ("N", "Isolation Status", 16),
        ("O", "Coverage Status", 18),
        ("P", "Gaps Identified", 30),
        ("Q", "Exception Approved", 16),
        ("R", "Target Date", 14),
        ("S", "Responsible Party", 20),
        ("T", "Notes", 25),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (row 7)
    example_data_net = [
        "Production LAN", "Internal LAN", "10.0.1.0/24", "VLAN-10", "45",
        "Critical", "Confidential", "Yes", "Yes", "Yes", "Yes (EDR)", "Yes",
        "Yes", "Active", "\u2705 Full", "None", "N/A", "N/A", "SOC Team", "Core production segment"
    ]
    row = 7
    for col_idx, value in enumerate(example_data_net, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    # Create extra validations outside the loop (not per-cell)
    edr_validation = DataValidation(type="list", formula1='"Yes (EDR),Partial,No"', allow_blank=False)
    tap_validation = DataValidation(type="list", formula1='"Yes,No,Planned"', allow_blank=False)
    net_cov_validation = DataValidation(type="list", formula1='"\u2705 Full,\u26A0\uFE0F Partial,\u274C None,N/A"', allow_blank=False)

    for data_row in range(8, 58):
        for col_idx in range(1, 21):  # A to T
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':
                validations['segment_type'].add(cell)
            elif col_letter == 'F':
                validations['criticality'].add(cell)
            elif col_letter == 'G':
                validations['data_classification'].add(cell)
            elif col_letter == 'H':
                validations['perimeter_monitoring'].add(cell)
            elif col_letter in ['I', 'J', 'L']:
                validations['yes_partial_no'].add(cell)
            elif col_letter == 'K':
                edr_validation.add(cell)
            elif col_letter == 'M':
                tap_validation.add(cell)
            elif col_letter == 'N':
                validations['isolation_status'].add(cell)
            elif col_letter == 'O':
                net_cov_validation.add(cell)
            elif col_letter == 'Q':
                validations['yes_no_na'].add(cell)

    # Compliance Checklist
    row = 59
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "NETWORK COVERAGE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "All network segments inventoried",
        "All production segments monitored",
        "All DMZ segments monitored",
        "Critical segments have redundant monitoring",
        "Perimeter traffic monitored (firewall + IDS/IPS)",
        "Internal traffic monitored (flow data)",
        "DNS queries monitored",
        "Endpoints monitored (EDR/AV integration)",
        "East-west traffic visibility (lateral movement detection)",
        "Cloud network segments included",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:S{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:S{row}")

        ws[f"T{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"T{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"T{row}"])
        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(T61:T70,"\u2705 Compliant")&" / 10"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Finalize all validations
    all_dvs = list(validations.values()) + [edr_validation, tap_validation, net_cov_validation]
    for _dv in all_dvs:
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: SHEET 4 - USER & IDENTITY COVERAGE
# ============================================================================

def create_user_identity_coverage_sheet(ws, styles):
    """Create User & Identity Monitoring Coverage sheet."""
    ws.row_dimensions[1].height = 35
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:S1")
    ws["A1"] = "3. USER & IDENTITY MONITORING COVERAGE"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:S2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.2 - Assess user monitoring"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        ("A", "Identity System", 25),
        ("B", "System Type", 22),
        ("C", "User Count", 15),
        ("D", "Privileged Account Count", 20),
        ("E", "Service Account Count", 20),
        ("F", "Authentication Logs Collected", 22),
        ("G", "Authorisation Logs Collected", 22),
        ("H", "Password Change Logs", 20),
        ("I", "Privilege Escalation Logs", 22),
        ("J", "MFA Events Logged", 18),
        ("K", "SSO Events Logged", 18),
        ("L", "Failed Login Monitoring", 20),
        ("M", "After-Hours Access Monitoring", 22),
        ("N", "Geographic Anomaly Detection", 22),
        ("O", "User Behaviour Analytics", 22),
        ("P", "Privileged Access Monitoring", 22),
        ("Q", "Coverage Status", 18),
        ("R", "Gaps/Issues", 30),
        ("S", "Priority", 16),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (row 7)
    example_data_uid = [
        "Microsoft Entra ID", "Cloud IdP", "2500", "45", "120",
        "Yes", "Yes", "Yes", "Yes", "Yes", "Yes",
        "Yes", "Yes", "Yes", "Yes (UEBA)", "Yes (PAM integrated)",
        "\u2705 Compliant", "None", "High"
    ]
    row = 7
    for col_idx, value in enumerate(example_data_uid, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    # Create extra validations outside the loop
    geo_validation = DataValidation(type="list", formula1='"Yes,No,Planned"', allow_blank=False)
    ueba_validation = DataValidation(type="list", formula1='"Yes (UEBA),Planned,No"', allow_blank=False)
    pam_validation = DataValidation(type="list", formula1='"Yes (PAM integrated),Partial,No"', allow_blank=False)

    for data_row in range(8, 58):
        for col_idx in range(1, 20):  # A to S
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':
                validations['identity_system_type'].add(cell)
            elif col_letter in ['F', 'G']:
                validations['yes_partial_no'].add(cell)
            elif col_letter in ['H', 'I', 'L', 'M']:
                validations['yes_no'].add(cell)
            elif col_letter in ['J', 'K']:
                validations['yes_no_na'].add(cell)
            elif col_letter == 'N':
                geo_validation.add(cell)
            elif col_letter == 'O':
                ueba_validation.add(cell)
            elif col_letter == 'P':
                pam_validation.add(cell)
            elif col_letter == 'Q':
                validations['compliance_status'].add(cell)
            elif col_letter == 'S':
                validations['priority'].add(cell)

    # Compliance Checklist
    row = 59
    ws.merge_cells(f"A{row}:S{row}")
    ws[f"A{row}"] = "USER & IDENTITY COVERAGE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Primary identity system (AD/Microsoft Entra ID (formerly Azure AD)) monitored",
        "All authentication events logged",
        "Failed login attempts monitored and alerted",
        "Privileged account usage monitored",
        "Service account usage monitored",
        "Password changes logged",
        "Privilege escalation logged",
        "MFA events logged (if MFA deployed)",
        "SSO events logged (if SSO deployed)",
        "After-hours access monitored",
        "Geographic anomalies detected",
        "User behaviour baselines established",
        "Dormant accounts identified",
        "Privileged access sessions recorded (PAM)",
        "Identity correlation across systems",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:R{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:R{row}")

        ws[f"S{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"S{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"S{row}"])
        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(S61:S75,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Finalize all validations
    all_dvs = list(validations.values()) + [geo_validation, ueba_validation, pam_validation]
    for _dv in all_dvs:
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 7: SHEET 5 - APPLICATION & SERVICE COVERAGE
# ============================================================================

def create_application_coverage_sheet(ws, styles):
    """Create Application & Service Monitoring Coverage sheet."""
    ws.row_dimensions[1].height = 35
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:U1")
    ws["A1"] = "4. APPLICATION & SERVICE MONITORING COVERAGE"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:U2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.2 - Assess application coverage"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        ("A", "Application/Service Name", 28),
        ("B", "Application Type", 22),
        ("C", "Business Unit", 20),
        ("D", "Application Owner", 20),
        ("E", "Data Classification", 18),
        ("F", "Criticality", 15),
        ("G", "User Base", 18),
        ("H", "Application Logs Collected", 22),
        ("I", "API Logs Collected", 18),
        ("J", "Database Logs Collected", 22),
        ("K", "Error/Exception Logging", 20),
        ("L", "Transaction Logging", 18),
        ("M", "Access Control Logs", 20),
        ("N", "Data Export Monitoring", 20),
        ("O", "Performance Monitoring", 20),
        ("P", "WAF Integration", 16),
        ("Q", "APM Integration", 16),
        ("R", "Coverage Status", 18),
        ("S", "Gaps", 30),
        ("T", "Target Date", 14),
        ("U", "Priority", 16),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "Customer Portal", "Web Application", "Sales", "M. Johnson", "Confidential",
        "High", "5000 users", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes",
        "Yes", "Yes", "Yes", "Yes", "\u2705 Compliant", "None", "N/A", "High"
    ]

    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    for data_row in range(8, 58):
        for col_idx in range(1, 22):  # A to U
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':
                validations['application_type'].add(cell)
            elif col_letter == 'E':
                validations['data_classification'].add(cell)
            elif col_letter == 'F':
                validations['criticality'].add(cell)
            elif col_letter in ['H', 'J', 'K']:
                validations['yes_partial_no'].add(cell)
            elif col_letter in ['I', 'L', 'M', 'N', 'O', 'Q']:
                validations['yes_no'].add(cell)
            elif col_letter == 'P':
                validations['yes_no_na'].add(cell)
            elif col_letter == 'R':
                validations['compliance_status'].add(cell)
            elif col_letter == 'U':
                validations['priority'].add(cell)

    # Compliance Checklist
    row = 60
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "APPLICATION COVERAGE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "All critical applications monitored",
        ">80% of high-priority applications monitored",
        "Application error/exception logs collected",
        "API calls logged (if API-based)",
        "Database queries logged (if database-backed)",
        "User access logged",
        "Data exports/downloads monitored",
        "File uploads monitored",
        "Configuration changes logged",
        "WAF deployed for internet-facing apps",
        "APM integrated for performance visibility",
        "Cloud application logs collected (SaaS)",
        "Mobile app backend logs collected",
        "Third-party integrations monitored",
        "Application security events forwarded to SIEM",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:T{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:T{row}")

        ws[f"U{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"U{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"U{row}"])
        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(U62:U76,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Finalize all validations
    for _dv in validations.values():
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: SHEET 6 - COVERAGE GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create Coverage Gap Analysis sheet."""
    ws.row_dimensions[1].height = 35
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:R1")
    ws["A1"] = "5. COVERAGE GAP ANALYSIS"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:R2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1 - Document and remediate coverage gaps"
    apply_style(ws["A2"], styles["subheader"])

    ws["A3"] = "Are all monitoring coverage gaps identified, documented, and tracked to remediation?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:R3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    headers = [
        ("A", "Gap ID", 12),
        ("B", "Gap Category", 22),
        ("C", "Affected Asset/System", 28),
        ("D", "Gap Description", 35),
        ("E", "Business Impact", 30),
        ("F", "Risk Level", 15),
        ("G", "Root Cause", 30),
        ("H", "Exception Approved", 16),
        ("I", "Exception ID", 15),
        ("J", "Compensating Controls", 30),
        ("K", "Remediation Plan", 35),
        ("L", "Remediation Owner", 20),
        ("M", "Target Date", 14),
        ("N", "Budget Required", 15),
        ("O", "Status", 18),
        ("P", "Status Date", 14),
        ("Q", "Verification Method", 25),
        ("R", "Notes", 30),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "GAP-001", "Asset Not Monitored", "Legacy File Server", "No log collection configured",
        "Medium - non-critical legacy system", "Medium", "End-of-life system, no agent support",
        "Yes", "EXC-2024-015", "Network-level monitoring via firewall",
        "Migrate to new storage platform", "Infrastructure Team", "31.03.2025",
        "Yes", "In Progress", "15.12.2024", "Post-migration verification", "Budget approved Q1"
    ]

    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-47: 40 rows)
    validations = create_base_validations(ws)

    budget_validation = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=False)
    exception_validation = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=False)

    for data_row in range(8, 48):
        # Auto-generate Gap ID
        ws[f"A{data_row}"] = f'="GAP-"&TEXT(ROW()-7,"000")'
        ws[f"A{data_row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"A{data_row}"].border = border

        for col_idx in range(2, 19):  # B to R
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")

            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':
                validations['gap_category'].add(cell)
            elif col_letter == 'F':
                validations['priority'].add(cell)
            elif col_letter == 'H':
                exception_validation.add(cell)
            elif col_letter == 'N':
                budget_validation.add(cell)
            elif col_letter == 'O':
                validations['gap_status'].add(cell)

    # Gap Summary Statistics (Starting Row 50)
    row = 50
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "GAP SUMMARY STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"F{row}"] = "Count/Value"
    ws[f"J{row}"] = "Target"
    ws[f"M{row}"] = "Status"
    for col in ['A', 'F', 'J', 'M']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:E{row}")
    ws.merge_cells(f"F{row}:I{row}")
    ws.merge_cells(f"J{row}:L{row}")
    ws.merge_cells(f"M{row}:R{row}")

    statistics = [
        ("Total Gaps Identified", "=COUNTA(B8:B47)", "N/A", ""),
        ("Critical Gaps", '=COUNTIF(F8:F47,"Critical")', "0", ""),
        ("High Gaps", '=COUNTIF(F8:F47,"High")', "<5", ""),
        ("Open Gaps", '=COUNTIF(O8:O47,"Open")', "Minimize", ""),
        ("In Progress", '=COUNTIF(O8:O47,"In Progress")', "N/A", ""),
        ("Resolved Gaps", '=COUNTIF(O8:O47,"Resolved")', "Maximize", ""),
        ("Accepted Risks", '=COUNTIF(O8:O47,"Accepted")', "Document", ""),
        ("Overdue (Past Target Date)", '=SUMPRODUCT((O8:O47<>"Resolved")*(O8:O47<>"Accepted")*(M8:M47<TODAY())*(M8:M47<>""))', "0", ""),
    ]

    row += 1
    for metric, formula, target, status in statistics:
        ws[f"A{row}"] = metric
        ws[f"F{row}"] = formula
        ws[f"J{row}"] = target
        ws[f"M{row}"] = status
        ws[f"M{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"M{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:E{row}")
        ws.merge_cells(f"F{row}:I{row}")
        ws.merge_cells(f"J{row}:L{row}")
        ws.merge_cells(f"M{row}:R{row}")
        row += 1

    # Compliance Checklist
    row = 62
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "GAP MANAGEMENT CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "All coverage gaps identified",
        "Gaps categorised by type",
        "Risk level assessed for each gap",
        "Root cause analysis conducted",
        "Critical gaps remediated within 30 days",
        "High gaps remediated within 90 days",
        "Exceptions formally approved",
        "Compensating controls documented",
        "Remediation plans documented",
        "Gap remediation tracked",
        "Verification conducted post-remediation",
        "Trends analysed (recurring gaps?)",
        "Gaps reported to CISO monthly",
        "Improvement actions implemented",
        "Coverage metrics improving over time",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:Q{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:Q{row}")

        ws[f"R{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"R{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"R{row}"])
        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(R64:R78,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Finalize all validations
    all_dvs = list(validations.values()) + [budget_validation, exception_validation]
    for _dv in all_dvs:
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with TABLE 1/2/3 Gold Standard layout."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: A1:G1 merged 003366 header
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle
    ws['A2'] = f'Summary Dashboard — {CONTROL_NAME} | {GENERATED_DATE}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.freeze_panes = 'A4'

    # TABLE 1: COMPLIANCE OVERVIEW
    ws.merge_cells('A4:G4')
    ws['A4'] = 'TABLE 1: COMPLIANCE OVERVIEW'
    ws['A4'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A4'].border = border

    # Row 5: headers
    t1_headers = ['Assessment Area', 'Total Items', 'Compliant', 'Partial', 'Non-Compliant', 'N/A', 'Compliance %']
    for col_idx, header in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='000000')
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Rows 6-10: data rows
    ws.cell(row=6, column=1, value='1. Asset Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=6, column=2, value='=SUM(C6:F6)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=3, value='=COUNTIF(\'1. Asset Coverage\'!W61:W75,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=4, value='=COUNTIF(\'1. Asset Coverage\'!W61:W75,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=5, value='=COUNTIF(\'1. Asset Coverage\'!W61:W75,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=6, value='=COUNTIF(\'1. Asset Coverage\'!W61:W75,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7, value='=IF((B6-F6)=0,0,C6/(B6-F6))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=6, column=col_idx).border = border

    ws.cell(row=7, column=1, value='2. Network Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=7, column=2, value='=SUM(C7:F7)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=3, value='=COUNTIF(\'2. Network Coverage\'!T61:T70,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=4, value='=COUNTIF(\'2. Network Coverage\'!T61:T70,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=5, value='=COUNTIF(\'2. Network Coverage\'!T61:T70,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=6, value='=COUNTIF(\'2. Network Coverage\'!T61:T70,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7, value='=IF((B7-F7)=0,0,C7/(B7-F7))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=7, column=col_idx).border = border

    ws.cell(row=8, column=1, value='3. User Identity Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=8, column=2, value='=SUM(C8:F8)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=3, value='=COUNTIF(\'3. User Identity Coverage\'!S61:S75,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=4, value='=COUNTIF(\'3. User Identity Coverage\'!S61:S75,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=5, value='=COUNTIF(\'3. User Identity Coverage\'!S61:S75,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=6, value='=COUNTIF(\'3. User Identity Coverage\'!S61:S75,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7, value='=IF((B8-F8)=0,0,C8/(B8-F8))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=8, column=col_idx).border = border

    ws.cell(row=9, column=1, value='4. Application Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=9, column=2, value='=SUM(C9:F9)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=3, value='=COUNTIF(\'4. Application Coverage\'!U62:U76,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=4, value='=COUNTIF(\'4. Application Coverage\'!U62:U76,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=5, value='=COUNTIF(\'4. Application Coverage\'!U62:U76,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=6, value='=COUNTIF(\'4. Application Coverage\'!U62:U76,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7, value='=IF((B9-F9)=0,0,C9/(B9-F9))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=9, column=col_idx).border = border

    ws.cell(row=10, column=1, value='5. Gap Analysis').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=10, column=2, value='=SUM(C10:F10)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=3, value='=COUNTIF(\'5. Gap Analysis\'!R64:R78,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=4, value='=COUNTIF(\'5. Gap Analysis\'!R64:R78,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=5, value='=COUNTIF(\'5. Gap Analysis\'!R64:R78,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=6, value='=COUNTIF(\'5. Gap Analysis\'!R64:R78,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7, value='=IF((B10-F10)=0,0,C10/(B10-F10))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=10, column=col_idx).border = border

    # Row 11: TOTAL
    ws.cell(row=11, column=1, value='TOTAL').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=1).border = border
    for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F'], start=2):
        ws.cell(row=11, column=col_idx, value=f'=SUM({col_letter}6:{col_letter}10)').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=11, column=col_idx).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws.cell(row=11, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=11, column=col_idx).border = border
    ws.cell(row=11, column=7, value='=IF((B11-F11)=0,0,C11/(B11-F11))').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=7).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=11, column=7).number_format = '0.0%'
    ws.cell(row=11, column=7).border = border

    # TABLE 2: KEY METRICS
    ws.merge_cells('A13:G13')
    ws['A13'] = 'TABLE 2: KEY METRICS'
    ws['A13'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A13'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A13'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A13'].border = border
    ws.merge_cells('A14:E14')
    ws['A14'] = 'Metric'
    ws['A14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A14'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A14'].border = border
    ws.merge_cells('F14:G14')
    ws['F14'] = 'Value'
    ws['F14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F14'].border = border
    yllw_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws.merge_cells('A15:E15')
    ws.cell(row=15, column=1, value='Assets — Full Monitoring Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=15, column=1).border = border
    ws.merge_cells('F15:G15')
    ws.cell(row=15, column=6, value='=COUNTIF(\'1. Asset Coverage\'!R8:R200,"✅ Full Coverage")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=15, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=15, column=col_idx).border = border

    ws.merge_cells('A16:E16')
    ws.cell(row=16, column=1, value='Assets — No Monitoring Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=16, column=1).border = border
    ws.merge_cells('F16:G16')
    ws.cell(row=16, column=6, value='=COUNTIF(\'1. Asset Coverage\'!R8:R200,"❌ No Coverage")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=16, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=16, column=col_idx).border = border

    ws.merge_cells('A17:E17')
    ws.cell(row=17, column=1, value='Critical Assets — Not Monitored').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=17, column=1).border = border
    ws.merge_cells('F17:G17')
    ws.cell(row=17, column=6, value='=COUNTIFS(\'1. Asset Coverage\'!I8:I200,"Critical",\'1. Asset Coverage\'!R8:R200,"❌ No Coverage")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=17, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=17, column=col_idx).border = border

    ws.merge_cells('A18:E18')
    ws.cell(row=18, column=1, value='High Assets — Not Monitored').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=18, column=1).border = border
    ws.merge_cells('F18:G18')
    ws.cell(row=18, column=6, value='=COUNTIFS(\'1. Asset Coverage\'!I8:I200,"High",\'1. Asset Coverage\'!R8:R200,"❌ No Coverage")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=18, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=18, column=col_idx).border = border

    ws.merge_cells('A19:E19')
    ws.cell(row=19, column=1, value='Network Segments — Full Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=19, column=1).border = border
    ws.merge_cells('F19:G19')
    ws.cell(row=19, column=6, value='=COUNTIF(\'2. Network Coverage\'!O8:O200,"✅ Full")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=19, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=19, column=col_idx).border = border

    ws.merge_cells('A20:E20')
    ws.cell(row=20, column=1, value='Network Segments — No Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=20, column=1).border = border
    ws.merge_cells('F20:G20')
    ws.cell(row=20, column=6, value='=COUNTIF(\'2. Network Coverage\'!O8:O200,"❌ None")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=20, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=20, column=col_idx).border = border

    ws.merge_cells('A21:E21')
    ws.cell(row=21, column=1, value='Gaps Identified (Gap Analysis)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=21, column=1).border = border
    ws.merge_cells('F21:G21')
    ws.cell(row=21, column=6, value="=COUNTA('5. Gap Analysis'!B8:B200)").font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=21, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=21, column=col_idx).border = border

    ws.merge_cells('A22:E22')
    ws.cell(row=22, column=1, value='Mandatory Monitoring — Not Implemented').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=22, column=1).border = border
    ws.merge_cells('F22:G22')
    ws.cell(row=22, column=6, value='=COUNTIFS(\'1. Asset Coverage\'!K8:K200,"Mandatory",\'1. Asset Coverage\'!L8:L200,"No")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=22, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=22, column=col_idx).border = border

    # TABLE 3: CRITICAL FINDINGS
    ws.merge_cells('A24:G24')
    ws['A24'] = 'TABLE 3: CRITICAL FINDINGS'
    ws['A24'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A24'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A24'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A24'].border = border
    ws.merge_cells('A25:E25')
    ws['A25'] = 'Category / Finding'
    ws['A25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A25'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A25'].border = border
    ws['F25'] = 'Count'
    ws['F25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F25'].border = border
    ws['G25'] = 'Action'
    ws['G25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['G25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['G25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G25'].border = border

    ws.merge_cells('A26:E26')
    ws.cell(row=26, column=1, value='Assets — No Monitoring Coverage').fill = yllw_fill
    ws.cell(row=26, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=26, column=1).border = border
    ws.cell(row=26, column=6, value='=COUNTIF(\'1. Asset Coverage\'!R8:R200,"❌ No Coverage")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=26, column=6).fill = yllw_fill
    ws.cell(row=26, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=26, column=6).border = border
    ws.cell(row=26, column=7, value='').fill = yllw_fill
    ws.cell(row=26, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=26, column=col_idx).fill = yllw_fill
        ws.cell(row=26, column=col_idx).border = border

    ws.merge_cells('A27:E27')
    ws.cell(row=27, column=1, value='Critical Assets — Not Monitored').fill = yllw_fill
    ws.cell(row=27, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=27, column=1).border = border
    ws.cell(row=27, column=6, value='=COUNTIFS(\'1. Asset Coverage\'!I8:I200,"Critical",\'1. Asset Coverage\'!R8:R200,"❌ No Coverage")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=27, column=6).fill = yllw_fill
    ws.cell(row=27, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=27, column=6).border = border
    ws.cell(row=27, column=7, value='').fill = yllw_fill
    ws.cell(row=27, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=27, column=col_idx).fill = yllw_fill
        ws.cell(row=27, column=col_idx).border = border

    ws.merge_cells('A28:E28')
    ws.cell(row=28, column=1, value='High Assets — Not Monitored').fill = yllw_fill
    ws.cell(row=28, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=28, column=1).border = border
    ws.cell(row=28, column=6, value='=COUNTIFS(\'1. Asset Coverage\'!I8:I200,"High",\'1. Asset Coverage\'!R8:R200,"❌ No Coverage")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=28, column=6).fill = yllw_fill
    ws.cell(row=28, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=28, column=6).border = border
    ws.cell(row=28, column=7, value='').fill = yllw_fill
    ws.cell(row=28, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=28, column=col_idx).fill = yllw_fill
        ws.cell(row=28, column=col_idx).border = border

    ws.merge_cells('A29:E29')
    ws.cell(row=29, column=1, value='Network Segments — No Coverage').fill = yllw_fill
    ws.cell(row=29, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=29, column=1).border = border
    ws.cell(row=29, column=6, value='=COUNTIF(\'2. Network Coverage\'!O8:O200,"❌ None")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=29, column=6).fill = yllw_fill
    ws.cell(row=29, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=29, column=6).border = border
    ws.cell(row=29, column=7, value='').fill = yllw_fill
    ws.cell(row=29, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=29, column=col_idx).fill = yllw_fill
        ws.cell(row=29, column=col_idx).border = border

    ws.merge_cells('A30:E30')
    ws.cell(row=30, column=1, value='Mandatory Monitoring Not Done').fill = yllw_fill
    ws.cell(row=30, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=30, column=1).border = border
    ws.cell(row=30, column=6, value='=COUNTIFS(\'1. Asset Coverage\'!K8:K200,"Mandatory",\'1. Asset Coverage\'!L8:L200,"No")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=30, column=6).fill = yllw_fill
    ws.cell(row=30, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=30, column=6).border = border
    ws.cell(row=30, column=7, value='').fill = yllw_fill
    ws.cell(row=30, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=30, column=col_idx).fill = yllw_fill
        ws.cell(row=30, column=col_idx).border = border

    # Row 31: instruction
    ws.merge_cells('A31:G31')
    ws['A31'] = 'Filter source sheets using AutoFilter — see column headers above'
    ws['A31'].font = Font(name='Calibri', size=9, italic=True, color='003366')
    ws['A31'].alignment = Alignment(horizontal='left', vertical='center')

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — Gold Standard 8-column format."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: merged title, 003366 fill, white bold 14pt + borders all 8 cells
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _c in range(1, 9):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: generic subtitle — italic, left-aligned
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty separator

    # Row 4: column headers — 003366 fill, white bold
    er_headers = [
        ('Evidence ID', 15),
        ('Assessment Area', 25),
        ('Evidence Type', 22),
        ('Description', 40),
        ('Location/Path', 45),
        ('Date Collected', 16),
        ('Collected By', 20),
        ('Verification Status', 22),
    ]
    for col_idx, (header, width) in enumerate(er_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Log Export,Documentation,Report,Network scan,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    ws.add_data_validation(status_dv)

    type_dv.add("C6:C105")
    status_dv.add("H6:H105")

    # Row 5: F2F2F2 grey sample row — fully populated example
    sample = [
        'EV-001',
        'Coverage Assessment',
        'Report',
        'Monitoring coverage assessment report — Q1 2024 all critical systems',
        '\\\\fileserver\\isms\\evidence\\monitoring\\coverage-assessment-q1-2024.pdf',
        '2024-03-01',
        'Thomas Weber',
        'Verified',
    ]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.font = Font(name="Calibri", size=10)

    # Rows 6-105: 100 empty FFFFCC data rows (no pre-populated IDs)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Freeze pane at A5 (freeze rows 1–4)
    ws.freeze_panes = "A5"

def _as3_border():
    """Return a fresh thin border object."""
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet (golden standard)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _bm3(ref):
        ws[ref].border = _as3_border()

    # Row 1: Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Assessment approval workflow for {DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border
    # Row 3: COMPLETED BY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    # Completed By fields
    row += 1
    for label in ["Name:", "Title:", "Date:", "Signature:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        _bm3(f"B{row}")
        row += 1

    # 2 more approver sections
    approvers = [
        ("REVIEWED BY", "4472C4"),
        ("APPROVED BY", "003366"),
    ]

    row += 1  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        # 5 fields per approver: Name, Title, Date, Signature, Comments
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            _bm3(f"B{row}")
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "FINAL DECISION"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1

    ws[f"A{row}"] = "Decision:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    _bm3(f"B{row}")

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    dv_decision.add(ws[f"B{row}"])

    # NEXT REVIEW
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        _bm3(f"B{row}")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Finalize validations
    for dv in [dv_decision]:
        if dv.sqref:
            ws.add_data_validation(dv)

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 11: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "You can't protect what you can't see."
    
    Coverage gaps are blind spots. Blind spots are where attackers hide.
    This assessment turns on the lights in every dark corner of your estate.
    
    Assets without monitoring = Assets without protection.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.16.3 - Coverage Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities")
    logger.info("=" * 78)
    logger.info("\n>>> Systems Engineering: Comprehensive Coverage Assessment")
    logger.info(">>> Complete Visibility: Assets, Networks, Users, Applications")
    logger.info(">>> Audit-Ready: 70 compliance checkpoints, gap tracking")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("\u2705 Workbook created with 9 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")
    
    logger.info("  [1/9] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  \u2705 Instructions complete")

    logger.info("  [2/9] Creating Asset Inventory & Coverage...")
    create_asset_coverage_sheet(wb["1. Asset Coverage"], styles)
    logger.info("  \u2705 Asset coverage complete (43 asset rows, 15 checks)")

    logger.info("  [3/9] Creating Network Segment Coverage...")
    create_network_coverage_sheet(wb["2. Network Coverage"], styles)
    logger.info("  \u2705 Network coverage complete (25 segment rows, 10 checks)")

    logger.info("  [4/9] Creating User & Identity Coverage...")
    create_user_identity_coverage_sheet(wb["3. User Identity Coverage"], styles)
    logger.info("  \u2705 Identity coverage complete (15 system rows, 15 checks)")

    logger.info("  [5/9] Creating Application & Service Coverage...")
    create_application_coverage_sheet(wb["4. Application Coverage"], styles)
    logger.info("  \u2705 Application coverage complete (25 app rows, 15 checks)")

    logger.info("  [6/9] Creating Coverage Gap Analysis...")
    create_gap_analysis_sheet(wb["5. Gap Analysis"], styles)
    logger.info("  \u2705 Gap analysis complete (40 gap rows, 15 checks)")

    logger.info("  [8/9] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  \u2705 Evidence register complete (100 evidence rows)")

    logger.info("  [7/9] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  \u2705 Dashboard complete (consolidated coverage metrics)")

    logger.info("  [9/9] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  \u2705 Approval workflow complete (4-level sign-off)")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.16.3_Coverage_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"\u2705 SUCCESS: {output_path}")
    except Exception as e:
        logger.error(f"\u274C ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets:")
    logger.info("  \u2022 Instructions & Legend (usage guidance)")
    logger.info("  \u2022 1. Asset Coverage (43 asset rows, criticality tracking)")
    logger.info("  \u2022 2. Network Coverage (25 segment rows, perimeter/flow monitoring)")
    logger.info("  \u2022 3. User/Identity Coverage (15 identity system rows)")
    logger.info("  \u2022 4. Application Coverage (25 application rows)")
    logger.info("  \u2022 5. Gap Analysis (40 gap tracking rows)")
    logger.info("\n>>> Consolidation & Governance:")
    logger.info("  \u2022 Summary Dashboard (overall coverage metrics)")
    logger.info("  \u2022 Evidence Register (100 evidence entries)")
    logger.info("  \u2022 Approval Sign-Off (4-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info(">>> ASSESSMENT CAPABILITIES:")
    logger.info("  \u2022 70 compliance checkpoint items across 5 assessment areas")
    logger.info("  \u2022 43 asset inventory rows with coverage tracking")
    logger.info("  \u2022 25 network segment rows")
    logger.info("  \u2022 15 identity system rows")
    logger.info("  \u2022 25 application monitoring rows")
    logger.info("  \u2022 40 gap tracking rows with remediation plans")
    logger.info("  \u2022 Coverage by criticality (Critical/High/Medium/Low)")
    logger.info("  \u2022 Coverage by regulatory scope (PCI-DSS, HIPAA, GDPR, SOX)")
    logger.info("  \u2022 Automated compliance % calculations")
    logger.info("\n" + "─" * 78)
    logger.info(">>> KEY FEATURES:")
    logger.info("  \u2705 Comprehensive asset inventory with monitoring status")
    logger.info("  \u2705 Network segment coverage (production, DMZ, internal)")
    logger.info("  \u2705 User and identity system monitoring assessment")
    logger.info("  \u2705 Application-level coverage tracking")
    logger.info("  \u2705 Gap identification and remediation tracking")
    logger.info("  \u2705 Coverage requirements by criticality")
    logger.info("  \u2705 Regulatory scope compliance (PCI, HIPAA, GDPR, SOX)")
    logger.info("  \u2705 Exception management and compensating controls")
    logger.info("  \u2705 Multi-level approval workflow")
    logger.info("  \u2705 Quarterly review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(">>> NEXT STEPS:")
    logger.info("  1. Complete comprehensive asset inventory")
    logger.info("  2. Document all network segments")
    logger.info("  3. Assess identity system monitoring")
    logger.info("  4. Inventory application monitoring")
    logger.info("  5. Identify ALL coverage gaps")
    logger.info("  6. Prioritize gaps by criticality and regulatory requirements")
    logger.info("  7. Document remediation plans")
    logger.info("  8. Review Summary Dashboard")
    logger.info("  9. Gather evidence")
    logger.info("  10. Obtain approvals")
    logger.info("\n>>> PRO TIP:")
    logger.info("  Coverage assessment isn't a one-time exercise.")
    logger.info("  Your asset inventory changes. New systems deploy.")
    logger.info("  Applications evolve. Networks expand.")
    logger.info("  Run this quarterly. Track coverage over time.")
    logger.info("  Trend toward 100% coverage of Critical/High assets.")
    logger.info("  That's how you eliminate blind spots systematically.")
    logger.info("\n" + "=" * 78)
    logger.info('\n"You can\'t protect what you can\'t see."')
    logger.info("\n>>> This is not cargo cult ISMS. This is comprehensive visibility.")
    logger.info(">>> We inventory EVERYTHING. We identify EVERY gap. We track to closure.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
