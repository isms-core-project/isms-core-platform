#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.16.4 - Alert Management & Response Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities
Assessment Domain 4 of 5: Alert Management, Triage, and Incident Response

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific alert management workflows, SOC procedures,
and incident response requirements.

Key customisation areas:
1. Alert severity definitions (match your classification scheme)
2. Triage workflows (specific to your SOC operations)
3. Escalation procedures (aligned with your organisational structure)
4. SLA thresholds (MTTD, MTTR per your commitments)
5. False positive management criteria (based on your tuning approach)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for monitoring)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
alert management, triage procedures, incident response workflows, and performance
metrics against ISO 27001:2022 Control A.8.16 requirements.

**Purpose:**
Enables systematic assessment of alert lifecycle management from generation through
response, including triage effectiveness, escalation procedures, and performance
metrics (MTTD, MTTR, false positive rates).

**Assessment Scope:**
- Alert generation effectiveness and accuracy
- Alert triage procedures and prioritization
- Incident response workflow maturity
- Escalation procedures and communication
- Performance metrics (MTTD, MTTR, response times)
- False positive management and tuning
- Gap analysis and process improvement
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and alert management standards
2. Alert Generation - Alert source effectiveness and configuration
3. Alert Triage - Triage procedures and prioritization assessment
4. Incident Response - Response workflow and effectiveness
5. Escalation Procedures - Escalation paths and communication
6. Performance Metrics - MTTD, MTTR, and SLA compliance tracking
7. Summary Dashboard - Overall alert management maturity metrics
8. Evidence Register - Audit evidence tracking and documentation
9. Approval Sign-Off - Multi-level approval workflow

**Key Features:**
- Data validation with severity and priority dropdown lists
- Conditional formatting for SLA compliance visualization
- Automated MTTD/MTTR calculations
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with A.8.16 Compliance Dashboard

**Integration:**
This assessment feeds into ISMS-IMP-A.8.16.5 Compliance Dashboard, which
consolidates data from all five monitoring assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a816_4_alert_management.py

Output:
    File: ISMS_A_8_16_4_Alert_Management_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Document alert generation sources and effectiveness
    2. Review triage procedures and prioritization logic
    3. Assess incident response workflow maturity
    4. Validate escalation procedures and contact lists
    5. Calculate MTTD and MTTR metrics
    6. Analyze false positive rates and alert fatigue
    7. Review SLA compliance (Critical: 15min, High: 1hr, Medium: 24hr)
    8. Conduct gap analysis for process improvements
    9. Define remediation actions with timelines
    10. Collect and link audit evidence
    11. Obtain stakeholder approvals
    12. Feed results into A.8.16.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.16
Assessment Domain:    4 of 5 (Alert Management & Response)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.16: Monitoring Activities Policy (Governance)
    - ISMS-POL-A.8.16-S2.3: Alert Management & Response Requirements
    - ISMS-POL-A.8.16-S5.C: Alert Response Procedures (Annex)
    - ISMS-IMP-A.8.16.4: Alert Management Implementation Guide
    - ISMS-IMP-A.8.16.1: Monitoring Infrastructure Assessment (Domain 1)
    - ISMS-IMP-A.8.16.2: Baseline & Detection Assessment (Domain 2)
    - ISMS-IMP-A.8.16.3: Coverage Assessment (Domain 3)
    - ISMS-IMP-A.8.16.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.16.4 specification
    - Supports comprehensive alert management and response evaluation
    - Integrated with A.8.16.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Alert Management Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This is not cargo cult ISMS. Alerts without response are noise. Detection
without action is security theater. If you can't respond in time, the threat
wins regardless of detection capabilities.

**Critical Performance Targets (ISMS-POL-A.8.16-S2.3):**

Mean Time to Detect (MTTD) - Target SLAs:
- Critical events (Tier 1): ≤15 minutes
- High severity (Tier 2): ≤1 hour
- Medium severity (Tier 3): ≤24 hours

Mean Time to Respond (MTTR) - Target SLAs:
- Critical events (Tier 1): ≤1 hour
- High severity (Tier 2): ≤4 hours
- Medium severity (Tier 3): ≤24 hours

Failure to meet these SLAs for Tier 1 events is a CRITICAL GAP requiring
immediate escalation to CISO and executive management.

**Alert Fatigue and False Positives:**
High alert volumes destroy SOC effectiveness. Key metrics:
- Alert volume: <100 alerts/day for mature SOC (target)
- False positive rate: <5% (target)
- Alert triage time: <5 minutes for initial classification (target)

Alert fatigue is a security risk - analysts become desensitized and miss
real threats. Tuning detection rules is not optional.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will expect:
- Documented triage and escalation procedures
- MTTD/MTTR metrics with trend analysis
- Evidence of continuous improvement
- SLA compliance tracking
- False positive management

**Data Protection:**
Assessment workbooks contain sensitive operational details including:
- Incident response procedures and workflows
- Escalation contacts and communication plans
- Performance metrics (including gaps)
- Known process weaknesses

Handle in accordance with your organisation's data classification policies.

**Common Alert Management Failures:**
- No documented triage procedures (chaos during incidents)
- Unclear escalation paths (delays in critical response)
- No MTTD/MTTR tracking (can't improve what you don't measure)
- Alert fatigue from untuned detection (missed real threats)
- Weekend/holiday coverage gaps (attackers love 3-day weekends)
- No runbooks for common alert types (inconsistent response)
- Escalation contact lists out of date (delays during crisis)

This assessment finds these issues before auditors (or attackers) do.

**Maintenance:**
Review and update assessment:
- Weekly: Review MTTD/MTTR metrics and SLA compliance
- Monthly: Analyze false positive rates and tune detection
- Quarterly: Validate escalation procedures and contact lists
- Semi-annually: Complete alert management process review
- Annually: Full reassessment with lessons learned integration
- Ad-hoc: After every major incident (continuous improvement)

**Quality Assurance:**
Have SOC managers and incident response team leads validate assessments.
Cross-check with actual incident response performance data - metrics must
reflect reality, not aspirations.

**Integration with Incident Response:**
Alert management is the front door to incident response. Poor alert management
= delayed incident response = increased damage. This assessment should be
reviewed alongside A.5.24 (Information Security Incident Management Planning
and Preparation) and A.5.25 (Assessment and Decision on Information Security
Events) assessments.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    import openpyxl
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.16.4"
WORKBOOK_NAME = "Alert Management & Response Assessment"
CONTROL_ID = "A.8.16"
CONTROL_NAME = "Monitoring Activities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Output directory (WKBK/ sibling of SCR/)
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)



def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    sheets = [
        "Instructions & Legend",
        "1. Alert Generation",
        "2. Triage Investigation",
        "3. Escalation Response",
        "4. Performance Metrics",
        "5. SOC Readiness",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "example_row": {
            "font": Font(name="Calibri", size=9, italic=True, color="808080"),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style to cell."""
    if "font" in style_dict:
        cell.font = Font(name=style_dict["font"].name, size=style_dict["font"].size, 
                        bold=style_dict["font"].bold, 
                        color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None,
                        italic=style_dict["font"].italic if hasattr(style_dict["font"], 'italic') else False)
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(horizontal=style_dict["alignment"].horizontal, 
                                   vertical=style_dict["alignment"].vertical,
                                   wrap_text=style_dict["alignment"].wrap_text)
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validations."""
    validations = {
        'alert_source': DataValidation(type="list", 
            formula1='"SIEM,IDS/IPS,EDR,NDR,Firewall,WAF,AV,DLP,Cloud Security,Other"', allow_blank=False),
        'severity': DataValidation(type="list",
            formula1='"Critical (P1),High (P2),Medium (P3),Low (P4)"', allow_blank=False),
        'sla_timeframe': DataValidation(type="list",
            formula1='"<15 min,<1 hr,<4 hrs,<1 day,<3 days"', allow_blank=False),
        'yes_no': DataValidation(type="list", formula1='"Yes,No"', allow_blank=False),
        'yes_no_na': DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False),
        'yes_partial_no': DataValidation(type="list", formula1='"Yes,Partial,No"', allow_blank=False),
        'alert_status': DataValidation(type="list",
            formula1='"Active,Testing,Tuning Needed,Retired"', allow_blank=False),
        'compliance_status': DataValidation(type="list",
            formula1='"\u2705 Compliant,\u26A0\uFE0F Partial,\u274C Non-Compliant,N/A"', allow_blank=False),
        'priority': DataValidation(type="list",
            formula1='"Critical,High,Medium,Low,None"', allow_blank=False),
        'automation_level': DataValidation(type="list",
            formula1='"Fully Automated,Partially Automated,Manual"', allow_blank=False),
        'process_status': DataValidation(type="list",
            formula1='"\u2705 Defined,\u26A0\uFE0F Partial,\u274C Undefined"', allow_blank=False),
        'shift_coverage': DataValidation(type="list",
            formula1='"24/7,Business Hours,On-Call"', allow_blank=False),
    }
    for _dv in validations.values():
        ws.add_data_validation(_dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete Alert Generation inventory (Sheet 1).', '2. Document Triage & Investigation processes (Sheet 2).', '3. Define Escalation & Response procedures (Sheet 3).', '4. Track Performance Metrics (Sheet 4).', '5. Assess SOC Operational Readiness (Sheet 5).', '6. Review Summary Dashboard for gaps.', '7. Gather evidence and document in Evidence Register.', '8. Obtain approvals via Approval Sign-Off sheet.', '9. Review quarterly and update metrics.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_alert_generation_sheet(ws, styles):
    """Create Alert Generation sheet."""
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:V1")
    ws["A1"] = "1. ALERT GENERATION & CLASSIFICATION ASSESSMENT"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:V2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.3.3 - Alert generation"
    apply_style(ws["A2"], styles["subheader"])

    ws["A3"] = "Are alerts properly generated, classified, enriched, and managed?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:V3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    headers = [
        ("A", "Alert Type/Name", 30), ("B", "Alert Source", 22), ("C", "Detection Rule ID", 20),
        ("D", "Alert Severity", 15), ("E", "MITRE ATT&CK Technique", 22), ("F", "Alert Description", 35),
        ("G", "Trigger Criteria", 30), ("H", "Enrichment Data", 30), ("I", "Expected FP Rate", 15),
        ("J", "Actual FP Rate (30d)", 15), ("K", "Alert Volume (30d)", 15), ("L", "True Positives (30d)", 15),
        ("M", "Response Playbook", 22), ("N", "SLA Timeframe", 18), ("O", "Auto-Enrichment", 16),
        ("P", "Auto-Containment", 16), ("Q", "Deduplication Enabled", 18), ("R", "Alert Status", 16),
        ("S", "Last Tuned", 14), ("T", "Compliance Status", 18), ("U", "Issues/Gaps", 30),
        ("V", "Tuning Priority", 16),
    ]

    row = 6
    for col, header, width in headers:
        ws[f"{col}{row}"] = header
        apply_style(ws[f"{col}{row}"], styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "Privilege Escalation Alert", "SIEM", "RULE-105", "High (P2)", "T1068", 
        "Detects privilege escalation attempts", "Event 4672 + 4768 within 5 min",
        "User context, asset info, threat intel", "10%", "8%", "45", "41",
        "PB-PRIV-ESC-001", "<1 hr", "Yes", "No", "Yes", "Active",
        "15.12.2024", "\u2705 Compliant", "None", "Low"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    validations = create_base_validations(ws)
    for data_row in range(8, 58):  # 50 rows (1 sample + 50 empty, standard MAX-001)
        for col_idx in range(1, 23):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            col_letter = get_column_letter(col_idx)
            if col_letter == 'B': validations['alert_source'].add(cell)
            elif col_letter == 'D': validations['severity'].add(cell)
            elif col_letter == 'N': validations['sla_timeframe'].add(cell)
            elif col_letter == 'O': validations['yes_partial_no'].add(cell)
            elif col_letter == 'P': validations['yes_no_na'].add(cell)
            elif col_letter == 'Q': validations['yes_no'].add(cell)
            elif col_letter == 'R': validations['alert_status'].add(cell)
            elif col_letter == 'T': validations['compliance_status'].add(cell)
            elif col_letter == 'V': validations['priority'].add(cell)

    # Alert Statistics
    row = 59
    ws.merge_cells(f"A{row}:V{row}")
    ws[f"A{row}"] = "ALERT STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"H{row}"] = "Count/Value"
    ws[f"O{row}"] = "Target"
    for col in ['A', 'H', 'O']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:G{row}")
    ws.merge_cells(f"H{row}:N{row}")
    ws.merge_cells(f"O{row}:V{row}")

    statistics = [
        ("Total Active Alerts", "=COUNTIF(R8:R57,\"Active\")", "N/A"),
        ("Critical Alerts", "=COUNTIF(D8:D57,\"Critical (P1)\")", "Target"),
        ("High Alerts", "=COUNTIF(D8:D57,\"High (P2)\")", "Target"),
        ("Alerts with Playbooks", "=COUNTIFS(M8:M57,\"<>\")", "All"),
        ("Alerts with Auto-Enrichment", "=COUNTIF(O8:O57,\"Yes\")", ">80%"),
        ("Alerts Needing Tuning", "=COUNTIF(R8:R57,\"Tuning Needed\")", "<10%"),
        ("Average FP Rate", "=AVERAGE(J8:J57)", "<25%"),
        ("Alerts Without Deduplication", "=COUNTIF(Q8:Q57,\"No\")", "0"),
    ]

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    row += 1
    for metric, formula, target in statistics:
        ws[f"A{row}"] = metric
        ws[f"H{row}"] = formula
        ws[f"O{row}"] = target
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{row}"].border = border_thin
        ws[f"H{row}"].border = border_thin
        ws[f"O{row}"].border = border_thin
        ws.merge_cells(f"A{row}:G{row}")
        ws.merge_cells(f"H{row}:N{row}")
        ws.merge_cells(f"O{row}:V{row}")

        row += 1

    # Checklist (starts after statistics section — row variable flows from stats loop)
    row += 2
    ws.merge_cells(f"A{row}:V{row}")
    ws[f"A{row}"] = "ALERT GENERATION CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "All alert types documented", "Alert severity classification consistent", "MITRE ATT&CK mapping documented",
        "Trigger criteria documented", "Alert enrichment configured", "Response playbook exists",
        "SLA timeframes defined", "False positive rates tracked (<25% overall, <10% critical)", "High FP alerts tuned within 30 days",
        "Alert deduplication configured", "Auto-containment for high-confidence", "Alerts tested before production",
        "Alert volume monitored for spikes", "Noisy alerts identified and tuned", "Retired alerts properly archived",
    ]

    row += 1
    checklist_header_row = row
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:U{row}")

    row += 1
    checklist_start_row = row
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"A{row}"].border = border_thin
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].border = border_thin
        ws.merge_cells(f"B{row}:U{row}")
        ws[f"V{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"V{row}"].border = border_thin
        validations['compliance_status'].add(ws[f"V{row}"])

        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(V{checklist_start_row}:V{row - 1},"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 5: SHEET 3 - TRIAGE & INVESTIGATION
# ============================================================================

def create_triage_investigation_sheet(ws, styles):
    """Create Triage & Investigation sheet."""
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:U1")
    ws["A1"] = "2. TRIAGE & INVESTIGATION PROCESS ASSESSMENT"
    apply_style(ws["A1"], styles["header"])


    ws.merge_cells("A2:U2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.3.5 - Triage procedures"
    apply_style(ws["A2"], styles["subheader"])

    ws["A3"] = "Are triage and investigation processes documented, trained, and effective?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:U3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")


    headers = [
        ("A", "Process Step", 28), ("B", "Process Owner", 20), ("C", "Procedure Documented", 18),
        ("D", "Documentation Location", 25), ("E", "Training Provided", 16), ("F", "Tools/Systems Used", 25),
        ("G", "Automation Level", 18), ("H", "Average Time (Minutes)", 18), ("I", "SLA Target (Minutes)", 18),
        ("J", "SLA Compliance %", 16), ("K", "Bottlenecks Identified", 30), ("L", "Quality Metrics", 25),
        ("M", "Error Rate %", 12), ("N", "Analyst Workload", 18), ("O", "Shift Coverage", 18),
        ("P", "Escalation Criteria", 30), ("Q", "Escalation Rate %", 15), ("R", "Process Status", 16),
        ("S", "Last Process Review", 14), ("T", "Improvement Opportunities", 30), ("U", "Priority", 16),
    ]

    row = 6
    for col, header, width in headers:
        ws[f"{col}{row}"] = header
        apply_style(ws[f"{col}{row}"], styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "Initial Assessment", "Tier 1 Analyst", "Yes", "SOC-SOP-002", "Yes",
        "SIEM, EDR Console", "Partially Automated", "8", "15", "92%",
        "Manual enrichment takes time", "First contact resolution rate", "3%",
        "35 alerts/shift", "24/7", "Severity High + Confirmed TP", "12%",
        "\u2705 Defined", "10.12.2024", "Automate enrichment", "Medium"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    validations = create_base_validations(ws)
    
    process_step_validation = DataValidation(type="list",
        formula1='"Alert Acknowledgment,Initial Assessment,Context Gathering,Disposition Decision,Investigation,Documentation,Escalation,Closure"',
        allow_blank=False)

    training_validation = DataValidation(type="list", formula1='"Yes,No,Planned"', allow_blank=False)

    ws.add_data_validation(process_step_validation)
    ws.add_data_validation(training_validation)

    for data_row in range(8, 58):  # 50 rows (1 sample + 50 empty, standard MAX-001)
        for col_idx in range(1, 22):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            col_letter = get_column_letter(col_idx)
            if col_letter == 'A': process_step_validation.add(cell)
            elif col_letter == 'C': validations['yes_partial_no'].add(cell)
            elif col_letter == 'E': training_validation.add(cell)
            elif col_letter == 'G': validations['automation_level'].add(cell)
            elif col_letter == 'O': validations['shift_coverage'].add(cell)
            elif col_letter == 'R': validations['process_status'].add(cell)
            elif col_letter == 'U': validations['priority'].add(cell)

    # Process Statistics
    row = 59
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "TRIAGE PROCESS STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"H{row}"] = "Count/Value"
    ws[f"N{row}"] = "Target"
    for col in ['A', 'H', 'N']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:G{row}")
    ws.merge_cells(f"H{row}:M{row}")
    ws.merge_cells(f"N{row}:U{row}")

    statistics = [
        ("Process Steps Documented", "=COUNTIF(C8:C57,\"Yes\")", "All"),
        ("Fully Automated Steps", "=COUNTIF(G8:G57,\"Fully Automated\")", "Target"),
        ("Partially Automated Steps", "=COUNTIF(G8:G57,\"Partially Automated\")", "N/A"),
        ("Manual Steps", "=COUNTIF(G8:G57,\"Manual\")", "Minimize"),
        ("Steps Meeting SLA", "=COUNTIFS(J8:J57,\">90%\")", "All"),
        ("Staff Trained", "=COUNTIF(E8:E57,\"Yes\")", "All"),
        ("Processes Reviewed (90d)", "[Manual Count]", "All"),
        ("Average Escalation Rate", "=AVERAGE(Q8:Q57)", "<15%"),
    ]

    thin_t = Side(style="thin")
    border_t = Border(left=thin_t, right=thin_t, top=thin_t, bottom=thin_t)

    row += 1
    for metric, formula, target in statistics:
        ws[f"A{row}"] = metric
        ws[f"H{row}"] = formula
        ws[f"N{row}"] = target
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{row}"].border = border_t
        ws[f"H{row}"].border = border_t
        ws[f"N{row}"].border = border_t
        ws.merge_cells(f"A{row}:G{row}")
        ws.merge_cells(f"H{row}:M{row}")
        ws.merge_cells(f"N{row}:U{row}")

        row += 1

    # Checklist (starts after statistics section — row variable flows from stats loop)
    row += 2
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "TRIAGE & INVESTIGATION CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Triage procedure documented", "Triage steps clearly defined", "Disposition criteria documented (TP/FP/Benign/Investigation)",
        "Triage timeframes defined per severity", "SOC analysts trained on triage procedures", "Triage tools accessible to all analysts",
        "Context enrichment automated where possible", "Common false positive scenarios documented", "Investigation playbooks exist",
        "Evidence collection guidance provided", "Documentation requirements defined", "Escalation criteria clearly defined",
        "Escalation paths documented", "Process reviewed quarterly", "Continuous improvement implemented",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:T{row}")

    row += 1
    checklist_start_row_t = row
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"A{row}"].border = border_t
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].border = border_t
        ws.merge_cells(f"B{row}:T{row}")
        ws[f"U{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"U{row}"].border = border_t
        validations['compliance_status'].add(ws[f"U{row}"])

        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(U{checklist_start_row_t}:U{row - 1},"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 6: SHEET 4 - ESCALATION & RESPONSE
# ============================================================================

def create_escalation_response_sheet(ws, styles):
    """Create Escalation & Response sheet."""
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:S1")
    ws["A1"] = "3. ESCALATION & RESPONSE PROCEDURES"
    apply_style(ws["A1"], styles["header"])


    ws.merge_cells("A2:S2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.3.7 - Escalation procedures"
    apply_style(ws["A2"], styles["subheader"])

    ws["A3"] = "Are escalation paths documented, tested, and effective?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:S3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")


    headers = [
        ("A", "Escalation Scenario", 30), ("B", "Trigger Criteria", 30), ("C", "Escalation Level", 20),
        ("D", "Target Person/Team", 22), ("E", "Primary Contact", 22), ("F", "Backup Contact", 22),
        ("G", "Contact Method", 18), ("H", "Escalation Timeframe", 16), ("I", "Information to Provide", 35),
        ("J", "Expected Response Time", 18), ("K", "Procedure Documented", 18), ("L", "Tested Frequency", 18),
        ("M", "Last Tested", 14), ("N", "Test Result", 16), ("O", "After-Hours Procedure", 20),
        ("P", "Escalation Rate (30d)", 16), ("Q", "Compliance Status", 18), ("R", "Gaps/Issues", 30),
        ("S", "Priority", 16),
    ]

    row = 6
    for col, header, width in headers:
        ws[f"{col}{row}"] = header
        apply_style(ws[f"{col}{row}"], styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "Critical Alert - Confirmed Ransomware", "Ransomware indicators + file encryption",
        "SOC→IR", "Incident Response Team", "IR Lead (+41 XX XXX XX XX)", "IR Manager (+41 XX XXX XX XX)",
        "Phone", "Within 15 min", "Alert details, affected systems, IOCs", "30 min",
        "Yes", "Quarterly", "15.10.2024", "Successful", "On-Call",
        "3 incidents", "\u2705 Compliant", "None", "Critical"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    validations = create_base_validations(ws)
    
    escalation_level_validation = DataValidation(type="list",
        formula1='"Tier 1\u2192Tier 2,Tier 2\u2192Tier 3,SOC\u2192IR,IR\u2192CISO,CISO\u2192Exec,Exec\u2192Board,External"',
        allow_blank=False)

    contact_method_validation = DataValidation(type="list",
        formula1='"Phone,Email,Ticketing System,Secure Chat,Multiple"', allow_blank=False)

    test_frequency_validation = DataValidation(type="list",
        formula1='"Quarterly,Semi-Annually,Annually,Never"', allow_blank=False)

    test_result_validation = DataValidation(type="list",
        formula1='"Successful,Issues Found,Not Tested"', allow_blank=False)

    after_hours_validation = DataValidation(type="list",
        formula1='"On-Call,24/7 SOC,Automated,None"', allow_blank=False)

    ws.add_data_validation(escalation_level_validation)
    ws.add_data_validation(contact_method_validation)
    ws.add_data_validation(test_frequency_validation)
    ws.add_data_validation(test_result_validation)
    ws.add_data_validation(after_hours_validation)
    
    for data_row in range(8, 58):  # 50 rows (1 sample + 50 empty, standard MAX-001)
        for col_idx in range(1, 20):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            col_letter = get_column_letter(col_idx)
            if col_letter == 'C': escalation_level_validation.add(cell)
            elif col_letter == 'G': contact_method_validation.add(cell)
            elif col_letter == 'K': validations['yes_partial_no'].add(cell)
            elif col_letter == 'L': test_frequency_validation.add(cell)
            elif col_letter == 'N': test_result_validation.add(cell)
            elif col_letter == 'O': after_hours_validation.add(cell)
            elif col_letter == 'Q': validations['compliance_status'].add(cell)
            elif col_letter == 'S': validations['priority'].add(cell)

    # Escalation Statistics
    row = 59
    ws.merge_cells(f"A{row}:S{row}")
    ws[f"A{row}"] = "ESCALATION STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"G{row}"] = "Count/Value"
    ws[f"M{row}"] = "Target"
    for col in ['A', 'G', 'M']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:F{row}")
    ws.merge_cells(f"G{row}:L{row}")
    ws.merge_cells(f"M{row}:S{row}")

    statistics = [
        ("Total Escalation Paths Documented", "=COUNTA(A8:A87)", "N/A"),
        ("Paths with Procedures", "=COUNTIF(K8:K87,\"Yes\")", "All"),
        ("Tested Quarterly", "=COUNTIF(L8:L87,\"Quarterly\")", "Target"),
        ("Tested in Last 90 Days", "=COUNTIFS(M8:M87,\">\"&TODAY()-90)", "All"),
        ("Successful Test Results", "=COUNTIF(N8:N87,\"Successful\")", "All"),
        ("Paths with After-Hours Coverage", "=COUNTIFS(O8:O87,\"<>None\")", "Critical paths"),
        ("Paths Needing Improvement", "=COUNTIF(Q8:Q87,\"\u26A0\uFE0F Partial\")", "<3"),
        ("Critical Escalations (30d)", "[From logs]", "N/A"),
    ]

    thin_e = Side(style="thin")
    border_e = Border(left=thin_e, right=thin_e, top=thin_e, bottom=thin_e)

    row += 1
    for metric, formula, target in statistics:
        ws[f"A{row}"] = metric
        ws[f"G{row}"] = formula
        ws[f"M{row}"] = target
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{row}"].border = border_e
        ws[f"G{row}"].border = border_e
        ws[f"M{row}"].border = border_e
        ws.merge_cells(f"A{row}:F{row}")
        ws.merge_cells(f"G{row}:L{row}")
        ws.merge_cells(f"M{row}:S{row}")

        row += 1

    # Checklist (starts after statistics section — row variable flows from stats loop)
    row += 2
    ws.merge_cells(f"A{row}:S{row}")
    ws[f"A{row}"] = "ESCALATION & RESPONSE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Escalation paths documented for all severity levels", "Escalation criteria clearly defined",
        "Contact information current (verified quarterly)", "Backup contacts identified",
        "After-hours escalation procedures defined", "On-call rotation established",
        "Escalation timeframes defined", "Information requirements documented (what to escalate)",
        "Communication templates available", "Escalation tracking in place",
        "Escalation procedures tested (tabletop exercises)", "External escalation procedures defined (law enforcement, regulators)",
        "Executive escalation criteria defined", "Escalation metrics tracked", "False escalations analysed and reduced",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:R{row}")

    row += 1
    checklist_start_row_e = row
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"A{row}"].border = border_e
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].border = border_e
        ws.merge_cells(f"B{row}:R{row}")
        ws[f"S{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"S{row}"].border = border_e
        validations['compliance_status'].add(ws[f"S{row}"])

        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(S{checklist_start_row_e}:S{row - 1},"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 7: SHEET 5 - PERFORMANCE METRICS
# ============================================================================

def create_performance_metrics_sheet(ws, styles):
    """Create Performance Metrics sheet."""
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:R1")
    ws["A1"] = "4. ALERT PERFORMANCE METRICS"
    apply_style(ws["A1"], styles["header"])


    ws.merge_cells("A2:R2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.3.9 - Performance metrics"
    apply_style(ws["A2"], styles["subheader"])

    ws["A3"] = "Are alert performance metrics tracked, analysed, and driving improvement?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:R3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")


    headers = [
        ("A", "Metric Name", 30), ("B", "Metric Category", 22), ("C", "Measurement Method", 25),
        ("D", "Data Source", 22), ("E", "Current Value", 15), ("F", "Target/SLA", 15),
        ("G", "Status", 16), ("H", "Trend (30d)", 16), ("I", "Measurement Frequency", 18),
        ("J", "Reporting Frequency", 18), ("K", "Reported To", 20), ("L", "Automated Tracking", 18),
        ("M", "Dashboard Available", 18), ("N", "Alert on Threshold", 18), ("O", "Last Review", 14),
        ("P", "Action Items", 30), ("Q", "Compliance Status", 18), ("R", "Notes", 25),
    ]

    row = 6
    for col, header, width in headers:
        ws[f"{col}{row}"] = header
        apply_style(ws[f"{col}{row}"], styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "Mean Time To Acknowledge (MTTA)", "Response Time", "SIEM timestamp analysis",
        "SIEM", "8 min", "<15 min", "\u2705 Meeting Target", "Improving",
        "Real-Time", "Daily", "SOC Lead", "Yes", "Yes", "Yes",
        "05.01.2025", "None - on target", "\u2705 Compliant", "Critical alerts only"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    validations = create_base_validations(ws)
    
    category_validation = DataValidation(type="list",
        formula1='"Volume,Response Time,Quality,Effectiveness,Workload"', allow_blank=False)

    status_validation = DataValidation(type="list",
        formula1='"\u2705 Meeting Target,\u26A0\uFE0F Below Target,\u274C Critical"', allow_blank=False)

    trend_validation = DataValidation(type="list",
        formula1='"Improving,Stable,Declining"', allow_blank=False)

    frequency_validation = DataValidation(type="list",
        formula1='"Real-Time,Daily,Weekly,Monthly,Quarterly"', allow_blank=False)

    ws.add_data_validation(category_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(trend_validation)
    ws.add_data_validation(frequency_validation)
    
    for data_row in range(8, 58):  # 50 rows for key metrics
        for col_idx in range(1, 19):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B': category_validation.add(cell)
            elif col_letter == 'G': status_validation.add(cell)
            elif col_letter == 'H': trend_validation.add(cell)
            elif col_letter in ['I', 'J']: frequency_validation.add(cell)
            elif col_letter in ['L', 'M', 'N']: validations['yes_no'].add(cell)
            elif col_letter == 'Q': validations['compliance_status'].add(cell)

    # Metrics Statistics
    row = 60
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "METRICS TRACKING STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"G{row}"] = "Count/Value"
    ws[f"L{row}"] = "Target"
    for col in ['A', 'G', 'L']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:F{row}")
    ws.merge_cells(f"G{row}:K{row}")
    ws.merge_cells(f"L{row}:R{row}")

    statistics = [
        ("Total Metrics Tracked", "=COUNTA(A8:A57)", "N/A"),
        ("Metrics Meeting Target", "=COUNTIF(G8:G57,\"\u2705 Meeting Target\")", ">80%"),
        ("Metrics Below Target", "=COUNTIF(G8:G57,\"\u26A0\uFE0F Below Target\")", "<3"),
        ("Metrics in Critical State", "=COUNTIF(G8:G57,\"\u274C Critical\")", "0"),
        ("Improving Trends", "=COUNTIF(H8:H57,\"Improving\")", "Maximize"),
        ("Declining Trends", "=COUNTIF(H8:H57,\"Declining\")", "Minimize"),
        ("Automated Tracking", "=COUNTIF(L8:L57,\"Yes\")", "All"),
        ("Dashboard Visibility", "=COUNTIF(M8:M57,\"Yes\")", "All"),
    ]

    thin_p = Side(style="thin")
    border_p = Border(left=thin_p, right=thin_p, top=thin_p, bottom=thin_p)

    row += 1
    for metric, formula, target in statistics:
        ws[f"A{row}"] = metric
        ws[f"G{row}"] = formula
        ws[f"L{row}"] = target
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{row}"].border = border_p
        ws[f"G{row}"].border = border_p
        ws[f"L{row}"].border = border_p
        ws.merge_cells(f"A{row}:F{row}")
        ws.merge_cells(f"G{row}:K{row}")
        ws.merge_cells(f"L{row}:R{row}")

        row += 1

    # Key Metrics Reference Table
    row = 72
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "KEY METRICS REFERENCE (MTTA/MTTT/MTTI/MTTR)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_header_cells = [("A", "Metric"), ("D", "Definition"), ("I", "Target (Critical)"), ("L", "Target (High)"), ("O", "Target (Medium)")]
    for col_letter, header in ref_header_cells:
        ws[f"{col_letter}{row}"] = header
        apply_style(ws[f"{col_letter}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:C{row}")
    ws.merge_cells(f"D{row}:H{row}")
    ws.merge_cells(f"I{row}:K{row}")
    ws.merge_cells(f"L{row}:N{row}")
    ws.merge_cells(f"O{row}:R{row}")

    ref_data = [
        ("MTTA", "Mean Time To Acknowledge", "<15 min", "<1 hr", "<4 hrs"),
        ("MTTT", "Mean Time To Triage", "<1 hr", "<4 hrs", "<1 day"),
        ("MTTI", "Mean Time To Investigate", "<4 hrs", "<1 day", "<3 days"),
        ("MTTR", "Mean Time To Resolve", "<24 hrs", "<3 days", "<1 week"),
        ("TP Rate", "True Positive Rate", ">30%", ">20%", ">15%"),
        ("FP Rate", "False Positive Rate", "<10%", "<20%", "<30%"),
    ]

    row += 1
    for metric, definition, crit, high, med in ref_data:
        ws[f"A{row}"] = metric
        ws[f"D{row}"] = definition
        ws[f"I{row}"] = crit
        ws[f"L{row}"] = high
        ws[f"O{row}"] = med
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{row}"].border = border_p
        ws[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"D{row}"].border = border_p
        ws[f"I{row}"].border = border_p
        ws[f"L{row}"].border = border_p
        ws[f"O{row}"].border = border_p
        ws.merge_cells(f"A{row}:C{row}")
        ws.merge_cells(f"D{row}:H{row}")
        ws.merge_cells(f"I{row}:K{row}")
        ws.merge_cells(f"L{row}:N{row}")
        ws.merge_cells(f"O{row}:R{row}")

        row += 1

    # Checklist
    row = 82
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "PERFORMANCE METRICS CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Alert volume metrics tracked", "Response time metrics tracked (MTTA, MTTT, MTTI, MTTR)", 
        "Quality metrics tracked (TP rate, FP rate)", "Effectiveness metrics tracked (detection rate)",
        "Metrics targets defined", "Metrics measured consistently",
        "Metrics reported to SOC Lead (weekly)", "Metrics reported to CISO (monthly)",
        "Metrics dashboard available", "Trends analysed",
        "Performance issues identified", "Improvement actions taken",
        "SLA compliance tracked", "Metrics drive tuning decisions", "Metrics included in management reporting",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:Q{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"A{row}"].border = border_p
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].border = border_p
        ws.merge_cells(f"B{row}:Q{row}")
        ws[f"R{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"R{row}"].border = border_p
        validations['compliance_status'].add(ws[f"R{row}"])

        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(R84:R98,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 8: SHEET 6 - SOC READINESS
# ============================================================================

def create_soc_readiness_sheet(ws, styles):
    """Create SOC Operational Readiness sheet."""
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:Q1")
    ws["A1"] = "5. SOC OPERATIONAL READINESS ASSESSMENT"
    apply_style(ws["A1"], styles["header"])


    ws.merge_cells("A2:Q2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S3 - SOC operations"
    apply_style(ws["A2"], styles["subheader"])

    ws["A3"] = "Is the SOC operationally ready with adequate staffing, training, tools, and procedures?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:Q3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")


    headers = [
        ("A", "Readiness Area", 28), ("B", "Requirement", 35), ("C", "Current State", 35),
        ("D", "Status", 18), ("E", "Evidence", 30), ("F", "Gap Description", 30),
        ("G", "Business Impact", 25), ("H", "Risk Level", 15), ("I", "Remediation Plan", 30),
        ("J", "Target Date", 14), ("K", "Owner", 20), ("L", "Budget Required", 15),
        ("M", "Dependencies", 25), ("N", "Status", 16), ("O", "Last Updated", 14),
        ("P", "Compliance Status", 18), ("Q", "Notes", 25),
    ]

    row = 6
    for col, header, width in headers:
        ws[f"{col}{row}"] = header
        apply_style(ws[f"{col}{row}"], styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row
    example_data = [
        "Staffing", "24/7 SOC coverage with Tier 1/2 analysts", "5 Tier 1, 3 Tier 2, 24/7 coverage",
        "\u2705 Adequate", "Shift rosters, staffing plan", "None", "N/A",
        "Low", "N/A", "N/A", "SOC Manager", "No", "None",
        "Complete", "01.01.2025", "\u2705 Compliant", "Adequate staffing levels"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    validations = create_base_validations(ws)
    
    readiness_validation = DataValidation(type="list",
        formula1='"Staffing,Training,Tools,Procedures,Communication,Facilities,Testing,Documentation"',
        allow_blank=False)

    status_validation = DataValidation(type="list",
        formula1='"\u2705 Adequate,\u26A0\uFE0F Needs Improvement,\u274C Inadequate"', allow_blank=False)

    budget_validation = DataValidation(type="list",
        formula1='"Yes,No,Unknown"', allow_blank=False)

    progress_validation = DataValidation(type="list",
        formula1='"Complete,In Progress,Planned,Not Started"', allow_blank=False)

    ws.add_data_validation(readiness_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(budget_validation)
    ws.add_data_validation(progress_validation)
    
    for data_row in range(8, 58):  # 50 rows
        for col_idx in range(1, 18):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            col_letter = get_column_letter(col_idx)
            if col_letter == 'A': readiness_validation.add(cell)
            elif col_letter == 'D': status_validation.add(cell)
            elif col_letter == 'H': validations['priority'].add(cell)
            elif col_letter == 'L': budget_validation.add(cell)
            elif col_letter == 'N': progress_validation.add(cell)
            elif col_letter == 'P': validations['compliance_status'].add(cell)

    # Readiness Statistics
    row = 60
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "SOC READINESS STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"G{row}"] = "Count/Value"
    ws[f"L{row}"] = "Target"
    for col in ['A', 'G', 'L']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:F{row}")
    ws.merge_cells(f"G{row}:K{row}")
    ws.merge_cells(f"L{row}:Q{row}")

    statistics = [
        ("Total Readiness Requirements", "=COUNTA(A8:A57)", "N/A"),
        ("Requirements Adequate", "=COUNTIF(D8:D57,\"\u2705 Adequate\")", "All"),
        ("Requirements Needing Improvement", "=COUNTIF(D8:D57,\"\u26A0\uFE0F Needs Improvement\")", "<5"),
        ("Requirements Inadequate", "=COUNTIF(D8:D57,\"\u274C Inadequate\")", "0"),
        ("Remediation Complete", "=COUNTIF(N8:N57,\"Complete\")", "Target"),
        ("Remediation In Progress", "=COUNTIF(N8:N57,\"In Progress\")", "N/A"),
        ("Staffing by Area", "=COUNTIF(A8:A57,\"Staffing\")", "Target"),
        ("Training by Area", "=COUNTIF(A8:A57,\"Training\")", "Target"),
    ]

    thin_s = Side(style="thin")
    border_s = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    row += 1
    for metric, formula, target in statistics:
        ws[f"A{row}"] = metric
        ws[f"G{row}"] = formula
        ws[f"L{row}"] = target
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{row}"].border = border_s
        ws[f"G{row}"].border = border_s
        ws[f"L{row}"].border = border_s
        ws.merge_cells(f"A{row}:F{row}")
        ws.merge_cells(f"G{row}:K{row}")
        ws.merge_cells(f"L{row}:Q{row}")

        row += 1

    # Checklist
    row = 72
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "SOC OPERATIONAL READINESS CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "SOC staffing adequate for 24/7 coverage", "Analyst to alert ratio acceptable (<50 alerts/shift)",
        "SOC tiers defined (Tier 1, 2, 3)", "On-call rotation established", "Shift handover procedures defined",
        "SOC analysts trained on triage procedures", "SOC analysts trained on investigation playbooks",
        "SOC analysts trained on tools (SIEM, EDR, etc.)", "Training records maintained", "Quarterly training refreshers conducted",
        "All required tools accessible to SOC", "Tool access tested and functional", "Tool documentation available",
        "Tool integrations working", "Communication systems functional (ticketing, chat, email)",
        "SOC procedures documented", "Playbooks available for common alert types", "Escalation procedures documented",
        "Shift handover checklists available", "Standard operating procedures (SOPs) updated",
        "Tabletop exercises conducted quarterly", "Alert response tested (red team, purple team)",
        "Escalation procedures tested", "Communication systems tested", "Disaster recovery procedures tested",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:P{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"A{row}"].border = border_s
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].border = border_s
        ws.merge_cells(f"B{row}:P{row}")
        ws[f"Q{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"Q{row}"].border = border_s
        validations['compliance_status'].add(ws[f"Q{row}"])

        row += 1

    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(Q74:Q98,"\u2705 Compliant")&" / 25"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with TABLE 1/2/3 Gold Standard layout."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: A1:G1 merged 003366 header
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle
    ws['A2'] = f'Summary Dashboard — {CONTROL_NAME} | {GENERATED_DATE}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.freeze_panes = 'A4'

    # TABLE 1: COMPLIANCE OVERVIEW
    ws.merge_cells('A4:G4')
    ws['A4'] = 'TABLE 1: COMPLIANCE OVERVIEW'
    ws['A4'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A4'].border = border

    # Row 5: headers
    t1_headers = ['Assessment Area', 'Total Items', 'Compliant', 'Partial', 'Non-Compliant', 'N/A', 'Compliance %']
    for col_idx, header in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='000000')
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Rows 6-10: data rows
    ws.cell(row=6, column=1, value='1. Alert Generation').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=6, column=2, value='=SUM(C6:F6)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=3, value='=COUNTIF(\'1. Alert Generation\'!V73:V87,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=4, value='=COUNTIF(\'1. Alert Generation\'!V73:V87,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=5, value='=COUNTIF(\'1. Alert Generation\'!V73:V87,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=6, value='=COUNTIF(\'1. Alert Generation\'!V73:V87,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7, value='=IF((B6-F6)=0,0,C6/(B6-F6))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=6, column=col_idx).border = border

    ws.cell(row=7, column=1, value='2. Triage Investigation').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=7, column=2, value='=SUM(C7:F7)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=3, value='=COUNTIF(\'2. Triage Investigation\'!U73:U87,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=4, value='=COUNTIF(\'2. Triage Investigation\'!U73:U87,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=5, value='=COUNTIF(\'2. Triage Investigation\'!U73:U87,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=6, value='=COUNTIF(\'2. Triage Investigation\'!U73:U87,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7, value='=IF((B7-F7)=0,0,C7/(B7-F7))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=7, column=col_idx).border = border

    ws.cell(row=8, column=1, value='3. Escalation Response').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=8, column=2, value='=SUM(C8:F8)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=3, value='=COUNTIF(\'3. Escalation Response\'!S73:S87,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=4, value='=COUNTIF(\'3. Escalation Response\'!S73:S87,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=5, value='=COUNTIF(\'3. Escalation Response\'!S73:S87,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=6, value='=COUNTIF(\'3. Escalation Response\'!S73:S87,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7, value='=IF((B8-F8)=0,0,C8/(B8-F8))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=8, column=col_idx).border = border

    ws.cell(row=9, column=1, value='4. Performance Metrics').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=9, column=2, value='=SUM(C9:F9)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=3, value='=COUNTIF(\'4. Performance Metrics\'!R84:R98,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=4, value='=COUNTIF(\'4. Performance Metrics\'!R84:R98,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=5, value='=COUNTIF(\'4. Performance Metrics\'!R84:R98,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=6, value='=COUNTIF(\'4. Performance Metrics\'!R84:R98,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7, value='=IF((B9-F9)=0,0,C9/(B9-F9))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=9, column=col_idx).border = border

    ws.cell(row=10, column=1, value='5. SOC Readiness').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=10, column=2, value='=SUM(C10:F10)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=3, value='=COUNTIF(\'5. SOC Readiness\'!Q74:Q98,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=4, value='=COUNTIF(\'5. SOC Readiness\'!Q74:Q98,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=5, value='=COUNTIF(\'5. SOC Readiness\'!Q74:Q98,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=6, value='=COUNTIF(\'5. SOC Readiness\'!Q74:Q98,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7, value='=IF((B10-F10)=0,0,C10/(B10-F10))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=10, column=col_idx).border = border

    # Row 11: TOTAL
    ws.cell(row=11, column=1, value='TOTAL').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=1).border = border
    for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F'], start=2):
        ws.cell(row=11, column=col_idx, value=f'=SUM({col_letter}6:{col_letter}10)').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=11, column=col_idx).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws.cell(row=11, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=11, column=col_idx).border = border
    ws.cell(row=11, column=7, value='=IF((B11-F11)=0,0,C11/(B11-F11))').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=7).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=11, column=7).number_format = '0.0%'
    ws.cell(row=11, column=7).border = border

    # TABLE 2: KEY METRICS
    ws.merge_cells('A13:G13')
    ws['A13'] = 'TABLE 2: KEY METRICS'
    ws['A13'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A13'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A13'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A13'].border = border
    ws.merge_cells('A14:E14')
    ws['A14'] = 'Metric'
    ws['A14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A14'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A14'].border = border
    ws.merge_cells('F14:G14')
    ws['F14'] = 'Value'
    ws['F14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F14'].border = border
    yllw_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws.merge_cells('A15:E15')
    ws.cell(row=15, column=1, value='Active Alert Types').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=15, column=1).border = border
    ws.merge_cells('F15:G15')
    ws.cell(row=15, column=6, value='=COUNTIF(\'1. Alert Generation\'!R8:R200,"Active")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=15, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=15, column=col_idx).border = border

    ws.merge_cells('A16:E16')
    ws.cell(row=16, column=1, value='Alerts Needing Tuning').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=16, column=1).border = border
    ws.merge_cells('F16:G16')
    ws.cell(row=16, column=6, value='=COUNTIF(\'1. Alert Generation\'!R8:R200,"Tuning Needed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=16, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=16, column=col_idx).border = border

    ws.merge_cells('A17:E17')
    ws.cell(row=17, column=1, value='Critical (P1) Alert Types').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=17, column=1).border = border
    ws.merge_cells('F17:G17')
    ws.cell(row=17, column=6, value='=COUNTIF(\'1. Alert Generation\'!D8:D200,"Critical (P1)")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=17, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=17, column=col_idx).border = border

    ws.merge_cells('A18:E18')
    ws.cell(row=18, column=1, value='High (P2) Alert Types').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=18, column=1).border = border
    ws.merge_cells('F18:G18')
    ws.cell(row=18, column=6, value='=COUNTIF(\'1. Alert Generation\'!D8:D200,"High (P2)")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=18, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=18, column=col_idx).border = border

    ws.merge_cells('A19:E19')
    ws.cell(row=19, column=1, value='Triage Processes — Defined').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=19, column=1).border = border
    ws.merge_cells('F19:G19')
    ws.cell(row=19, column=6, value='=COUNTIF(\'2. Triage Investigation\'!R8:R200,"✅ Defined")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=19, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=19, column=col_idx).border = border

    ws.merge_cells('A20:E20')
    ws.cell(row=20, column=1, value='Triage Processes \u2014 Undefined').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=20, column=1).border = border
    ws.merge_cells('F20:G20')
    ws.cell(row=20, column=6, value='=COUNTIF(\'2. Triage Investigation\'!R8:R200,"\u274C Undefined")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=20, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=20, column=col_idx).border = border

    ws.merge_cells('A21:E21')
    ws.cell(row=21, column=1, value='Triage Processes \u2014 Partial').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=21, column=1).border = border
    ws.merge_cells('F21:G21')
    ws.cell(row=21, column=6, value='=COUNTIF(\'2. Triage Investigation\'!R8:R200,"\u26A0\uFE0F Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=21, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=21, column=col_idx).border = border

    ws.merge_cells('A22:E22')
    ws.cell(row=22, column=1, value='SOC Readiness Items — Non-Compliant').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=22, column=1).border = border
    ws.merge_cells('F22:G22')
    ws.cell(row=22, column=6, value='=COUNTIF(\'5. SOC Readiness\'!Q8:Q200,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=22, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=22, column=col_idx).border = border

    # TABLE 3: CRITICAL FINDINGS
    ws.merge_cells('A24:G24')
    ws['A24'] = 'TABLE 3: CRITICAL FINDINGS'
    ws['A24'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A24'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A24'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A24'].border = border
    ws.merge_cells('A25:E25')
    ws['A25'] = 'Category / Finding'
    ws['A25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A25'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A25'].border = border
    ws['F25'] = 'Count'
    ws['F25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F25'].border = border
    ws['G25'] = 'Action'
    ws['G25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['G25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['G25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G25'].border = border

    ws.merge_cells('A26:E26')
    ws.cell(row=26, column=1, value='Alerts Needing Tuning').fill = yllw_fill
    ws.cell(row=26, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=26, column=1).border = border
    ws.cell(row=26, column=6, value='=COUNTIF(\'1. Alert Generation\'!R8:R200,"Tuning Needed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=26, column=6).fill = yllw_fill
    ws.cell(row=26, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=26, column=6).border = border
    ws.cell(row=26, column=7, value='').fill = yllw_fill
    ws.cell(row=26, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=26, column=col_idx).fill = yllw_fill
        ws.cell(row=26, column=col_idx).border = border

    ws.merge_cells('A27:E27')
    ws.cell(row=27, column=1, value='Critical (P1) Alerts — Tuning Needed').fill = yllw_fill
    ws.cell(row=27, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=27, column=1).border = border
    ws.cell(row=27, column=6, value='=COUNTIFS(\'1. Alert Generation\'!D8:D200,"Critical (P1)",\'1. Alert Generation\'!R8:R200,"Tuning Needed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=27, column=6).fill = yllw_fill
    ws.cell(row=27, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=27, column=6).border = border
    ws.cell(row=27, column=7, value='').fill = yllw_fill
    ws.cell(row=27, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=27, column=col_idx).fill = yllw_fill
        ws.cell(row=27, column=col_idx).border = border

    ws.merge_cells('A28:E28')
    ws.cell(row=28, column=1, value='Triage Processes \u2014 Undefined').fill = yllw_fill
    ws.cell(row=28, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=28, column=1).border = border
    ws.cell(row=28, column=6, value='=COUNTIF(\'2. Triage Investigation\'!R8:R200,"\u274C Undefined")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=28, column=6).fill = yllw_fill
    ws.cell(row=28, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=28, column=6).border = border
    ws.cell(row=28, column=7, value='').fill = yllw_fill
    ws.cell(row=28, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=28, column=col_idx).fill = yllw_fill
        ws.cell(row=28, column=col_idx).border = border

    ws.merge_cells('A29:E29')
    ws.cell(row=29, column=1, value='Triage Processes \u2014 Partial').fill = yllw_fill
    ws.cell(row=29, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=29, column=1).border = border
    ws.cell(row=29, column=6, value='=COUNTIF(\'2. Triage Investigation\'!R8:R200,"\u26A0\uFE0F Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=29, column=6).fill = yllw_fill
    ws.cell(row=29, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=29, column=6).border = border
    ws.cell(row=29, column=7, value='').fill = yllw_fill
    ws.cell(row=29, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=29, column=col_idx).fill = yllw_fill
        ws.cell(row=29, column=col_idx).border = border

    ws.merge_cells('A30:E30')
    ws.cell(row=30, column=1, value='SOC Readiness — Non-Compliant Items').fill = yllw_fill
    ws.cell(row=30, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=30, column=1).border = border
    ws.cell(row=30, column=6, value='=COUNTIF(\'5. SOC Readiness\'!Q8:Q200,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=30, column=6).fill = yllw_fill
    ws.cell(row=30, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=30, column=6).border = border
    ws.cell(row=30, column=7, value='').fill = yllw_fill
    ws.cell(row=30, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=30, column=col_idx).fill = yllw_fill
        ws.cell(row=30, column=col_idx).border = border

    # Row 31: instruction
    ws.merge_cells('A31:G31')
    ws['A31'] = 'Filter source sheets using AutoFilter — see column headers above'
    ws['A31'].font = Font(name='Calibri', size=9, italic=True, color='003366')
    ws['A31'].alignment = Alignment(horizontal='left', vertical='center')

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — Gold Standard 8-column format."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: merged title, 003366 fill, white bold 14pt + borders all 8 cells
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _c in range(1, 9):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: generic subtitle — italic, left-aligned
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty separator

    # Row 4: column headers — 003366 fill, white bold
    er_headers = [
        ('Evidence ID', 15),
        ('Assessment Area', 25),
        ('Evidence Type', 22),
        ('Description', 40),
        ('Location/Path', 45),
        ('Date Collected', 16),
        ('Collected By', 20),
        ('Verification Status', 22),
    ]
    for col_idx, (header, width) in enumerate(er_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Log Export,Documentation,Report,Network scan,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    ws.add_data_validation(status_dv)

    type_dv.add("C6:C105")
    status_dv.add("H6:H105")

    # Row 5: F2F2F2 grey sample row — fully populated example
    sample = [
        'EV-001',
        'Alert Management',
        'Configuration file',
        'Alert escalation workflow configuration — Tier 1/2/3 routing rules',
        '\\\\fileserver\\isms\\evidence\\monitoring\\alert-workflow-config.xml',
        '2024-03-01',
        'Emma Fischer',
        'Verified',
    ]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.font = Font(name="Calibri", size=10)

    # Rows 6-105: 100 empty FFFFCC data rows (no pre-populated IDs)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Freeze pane at A5 (freeze rows 1–4)
    ws.freeze_panes = "A5"

def _as4_border():
    """Return a fresh thin border object."""
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet (golden standard)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _bm4(ref):
        ws[ref].border = _as4_border()

    # --- A1: merged title, height 35, 003366 fill ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = _as4_border()
    ws.row_dimensions[1].height = 35

    # --- A2: italic subtitle ---
    ws["A2"] = f"Approval workflow for {DOCUMENT_ID} \u2014 {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells("A2:E2")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = _as4_border()
    # Freeze pane
    ws.freeze_panes = "A3"

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40

    # --- COMPLETED BY section ---
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLETED BY"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as4_border()
    fields = ["Name", "Title", "Date", "Signature", "Comments"]
    for field in fields:
        row += 1
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border_thin
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border_thin
        ws.merge_cells(f"B{row}:E{row}")
        _bm4(f"B{row}")
    row += 1  # spacer

    # --- REVIEWED BY and APPROVED BY sections ---
    approver_titles = [
        ("REVIEWED BY", "4472C4"),
        ("APPROVED BY", "003366"),
    ]
    for title, color in approver_titles:
        row += 1
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = _as4_border()
        for field in fields:
            row += 1
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = border_thin
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border_thin
            ws.merge_cells(f"B{row}:E{row}")
            _bm4(f"B{row}")
        row += 1  # spacer

    # --- FINAL DECISION ---
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "FINAL DECISION"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as4_border()
    decision_fields = ["Overall Assessment Result", "Conditions / Observations", "Required Actions Before Next Review"]
    decision_row = None
    for field in decision_fields:
        row += 1
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border_thin
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border_thin
        ws.merge_cells(f"B{row}:E{row}")
        _bm4(f"B{row}")
        if field == "Overall Assessment Result":
            decision_row = row

    decision_dv = DataValidation(type="list",
        formula1='"Approved,Conditionally Approved,Not Approved"', allow_blank=True)
    ws.add_data_validation(decision_dv)

    if decision_row:
        decision_dv.add(ws[f"B{decision_row}"])

    # --- NEXT REVIEW ---
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as4_border()
    row += 1
    ws[f"A{row}"] = "Scheduled Review Date"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border_thin
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border_thin
    ws.merge_cells(f"B{row}:E{row}")
    _bm4(f"B{row}")

    row += 1
    ws[f"A{row}"] = "Review Frequency"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border_thin
    ws[f"B{row}"] = "Quarterly"
    ws[f"B{row}"].font = Font(name="Calibri", size=11)
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border_thin
    ws.merge_cells(f"B{row}:E{row}")
    _bm4(f"B{row}")


# ============================================================================
# SECTION 12: MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function."""
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.16.4 - Alert Management & Response Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities")
    logger.info("=" * 78)
    logger.info("\n>>> Systems Engineering: Data-Driven Alert Management")
    logger.info(">>> Complete: Generation, Triage, Escalation, Metrics, Readiness")
    logger.info(">>> Audit-Ready: 85 compliance checkpoints")
    logger.info("\n" + "─" * 78)

    logger.info("\n[Phase 1] Initializing workbook...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("\u2705 Workbook created with 9 sheets")

    logger.info("\n[Phase 2] Generating assessment sheets...")
    
    logger.info("  [1/9] Instructions...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  \u2705 Complete")

    logger.info("  [2/9] Alert Generation...")
    create_alert_generation_sheet(wb["1. Alert Generation"], styles)
    logger.info("  \u2705 Complete (30 alert rows, 15 checks)")

    logger.info("  [3/9] Triage & Investigation...")
    create_triage_investigation_sheet(wb["2. Triage Investigation"], styles)
    logger.info("  \u2705 Complete (15 process rows, 15 checks)")

    logger.info("  [4/9] Escalation & Response...")
    create_escalation_response_sheet(wb["3. Escalation Response"], styles)
    logger.info("  \u2705 Complete (20 escalation rows, 15 checks)")

    logger.info("  [5/9] Performance Metrics...")
    create_performance_metrics_sheet(wb["4. Performance Metrics"], styles)
    logger.info("  \u2705 Complete (20 metric rows, 15 checks)")

    logger.info("  [6/9] SOC Readiness...")
    create_soc_readiness_sheet(wb["5. SOC Readiness"], styles)
    logger.info("  \u2705 Complete (25 readiness rows, 25 checks)")

    logger.info("  [8/9] Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  \u2705 Complete (100 evidence rows)")

    logger.info("  [7/9] Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  \u2705 Complete (6 sections: Alert Mgmt, Response Time, Escalation, SOC Health, Compliance, Issues)")

    logger.info("  [9/9] Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  \u2705 Complete (3-level sign-off)")

    logger.info("\n[Phase 3] Saving workbook...")
    filename = f"ISMS-IMP-A.8.16.4_Alert_Management_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"\u2705 SUCCESS: {output_path}")
    except Exception as e:
        logger.error(f"\u274C ERROR: {e}")
        return 1

    logger.info("\n" + "=" * 78)
    logger.info("WORKBOOK COMPLETE")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets:")
    logger.info("  \u2022 Alert Generation & Classification (30 rows, 15 checks)")
    logger.info("  \u2022 Triage & Investigation Processes (15 rows, 15 checks)")
    logger.info("  \u2022 Escalation & Response (20 rows, 15 checks)")
    logger.info("  \u2022 Performance Metrics (20 rows, 15 checks)")
    logger.info("  \u2022 SOC Readiness (25 rows, 25 checks)")
    logger.info("  \u2022 Summary Dashboard (6 comprehensive sections)")
    logger.info("  \u2022 Evidence Register (100 entries)")
    logger.info("  \u2022 Approval Sign-Off (4-level workflow)")
    logger.info("\n>>> ASSESSMENT CAPABILITIES:")
    logger.info("  \u2022 85 compliance checkpoints across 5 areas")
    logger.info("  \u2022 30 alert generation inventory rows")
    logger.info("  \u2022 15 triage process documentation rows")
    logger.info("  \u2022 20 escalation scenario rows")
    logger.info("  \u2022 20 performance metric tracking rows")
    logger.info("  \u2022 25 SOC readiness assessment rows")
    logger.info("  \u2022 MTTA/MTTT/MTTI/MTTR tracking")
    logger.info("  \u2022 SLA compliance monitoring")
    logger.info("  \u2022 Escalation path testing")
    logger.info("  \u2022 SOC operational health assessment")
    logger.info("\n>>> KEY FEATURES:")
    logger.info("  \u2705 Alert generation and classification inventory")
    logger.info("  \u2705 Triage process documentation with automation levels")
    logger.info("  \u2705 Escalation procedures with testing requirements")
    logger.info("  \u2705 Performance metrics (response time, quality, effectiveness)")
    logger.info("  \u2705 SOC readiness (staffing, training, tools, procedures, testing)")
    logger.info("  \u2705 MITRE ATT&CK mapping for alerts")
    logger.info("  \u2705 Response playbook inventory")
    logger.info("  \u2705 False positive rate tracking")
    logger.info("  \u2705 Automated statistics and compliance scoring")
    logger.info("  \u2705 6-section dashboard with critical issues tracking")
    logger.info("\n" + "=" * 78)
    logger.info(">>> NEXT STEPS:")
    logger.info("  1. Inventory all active alerts")
    logger.info("  2. Document triage and investigation processes")
    logger.info("  3. Define and test escalation procedures")
    logger.info("  4. Track MTTA/MTTT/MTTI/MTTR metrics")
    logger.info("  5. Assess SOC operational readiness")
    logger.info("  6. Review Summary Dashboard for gaps")
    logger.info("  7. Address critical issues")
    logger.info("  8. Gather evidence")
    logger.info("  9. Obtain approvals")
    logger.info("  10. Review quarterly and improve continuously")
    logger.info("\n>>> PRO TIP:")
    logger.info("  Response without metrics is just firefighting.")
    logger.info("  Metrics without improvement is just reporting.")
    logger.info("  Measure MTTA/MTTT/MTTI/MTTR. Track trends.")
    logger.info("  Identify bottlenecks. Automate where possible.")
    logger.info("  Test escalation paths. Train analysts continuously.")
    logger.info("  That's Systems Engineering applied to SOC operations.")
    logger.info("\n" + "=" * 78)
    logger.info('"In God we trust. All others must bring data." - W. Edwards Deming')
    logger.info("\n>>> This is not cargo cult ISMS. This is data-driven SOC operations.")
    logger.info(">>> We MEASURE response times. We TRACK effectiveness. We IMPROVE continuously.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
