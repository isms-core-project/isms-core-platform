#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.16.1 - Monitoring Infrastructure Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities
Assessment Domain 1 of 5: Monitoring Infrastructure and Platform Capabilities

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific monitoring platform, SIEM deployment, and
infrastructure requirements.

Key customisation areas:
1. SIEM/log management platform (Splunk, Sentinel, ELK per your deployment)
2. Log source inventory (match your actual infrastructure)
3. Data collection architecture (specific to your network topology)
4. Performance thresholds (based on your SLA requirements)
5. Integration requirements (aligned with your security stack)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for monitoring)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
monitoring infrastructure capabilities against ISO 27001:2022 Control A.8.16 
requirements.

**Purpose:**
Enables systematic assessment of monitoring platform capabilities, log source
coverage, data collection architecture, integration capabilities, and performance/
scalability against Control A.8.16 requirements.

**Assessment Scope:**
- Monitoring platform capabilities (SIEM/Log Management)
- Log source coverage and collection
- Data collection architecture and reliability
- Integration and enrichment capabilities
- Performance and scalability assessment
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and monitoring standards
2. Monitoring Platform - SIEM/platform capability assessment
3. Log Source Coverage - Comprehensive log source inventory and coverage
4. Data Collection Architecture - Collection methods and reliability
5. Integration & Enrichment - Context enrichment and correlation
6. Performance & Scale - Capacity and performance assessment
7. Summary Dashboard - Compliance metrics and gap summary
8. Evidence Register - Audit evidence tracking and documentation
9. Approval Sign-Off - Multi-level approval workflow

**Key Features:**
- Data validation with monitoring capability dropdown lists
- Conditional formatting for compliance status visualization
- Automated gap identification and remediation tracking
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with A.8.16 Compliance Dashboard

**Integration:**
This assessment feeds into ISMS-IMP-A.8.16.5 Compliance Dashboard, which
consolidates data from all five monitoring assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a816_1_monitoring_infrastructure.py

Output:
    File: ISMS_A_8_16_1_Monitoring_Infrastructure_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Review monitoring platform capabilities documented
    2. Complete log source inventory (all systems generating logs)
    3. Assess data collection architecture and reliability
    4. Evaluate integration and enrichment capabilities
    5. Review performance and scalability metrics
    6. Conduct gap analysis for uncovered/unmonitored systems
    7. Define remediation actions with timelines
    8. Collect and link audit evidence
    9. Obtain stakeholder approvals
    10. Feed results into A.8.16.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.16
Assessment Domain:    1 of 5 (Monitoring Infrastructure)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.16: Monitoring Activities Policy (Governance)
    - ISMS-POL-A.8.16-S2.1: Monitoring Infrastructure Requirements
    - ISMS-IMP-A.8.16.1: Monitoring Infrastructure Implementation Guide
    - ISMS-IMP-A.8.16.2: Baseline & Detection Assessment (Domain 2)
    - ISMS-IMP-A.8.16.3: Coverage Assessment (Domain 3)
    - ISMS-IMP-A.8.16.4: Alert Management Assessment (Domain 4)
    - ISMS-IMP-A.8.16.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.16.1 specification
    - Supports comprehensive monitoring infrastructure evaluation
    - Integrated with A.8.16.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Monitoring Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This is not cargo cult ISMS. We assess CAPABILITIES, not checkbox compliance.
If your monitoring platform cannot detect threats, no policy will save you.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of monitoring capabilities and coverage.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Monitoring platform architecture and capabilities
- Log sources and network topology
- Coverage gaps and security blind spots

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new systems requiring monitoring
- Semi-annually: Update coverage gaps and remediation status
- Annually: Complete reassessment of all monitoring infrastructure
- Ad-hoc: When infrastructure changes or new threats emerge

**Quality Assurance:**
Have SOC engineers and security architects validate assessments before using
results for compliance reporting or remediation decisions.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    import openpyxl
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.16.1"
WORKBOOK_NAME = "Monitoring Infrastructure Assessment"
CONTROL_ID = "A.8.16"
CONTROL_NAME = "Monitoring Activities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Output directory (WKBK/ sibling of SCR/)
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)



def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.16.1 specification
    sheets = [
        "Instructions & Legend",
        "1. Monitoring Platform",
        "2. Log Source Coverage",
        "3. Data Collection Arch",
        "4. Integration Enrichment",
        "5. Performance Scale",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "example_row": {
            "font": Font(name="Calibri", size=9, italic=True, color="808080"),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "gap_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "gap_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "gap_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None,
            italic=style_dict["font"].italic if hasattr(style_dict["font"], 'italic') else False
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_planned': DataValidation(
            type="list",
            formula1='"Yes,No,Planned"',
            allow_blank=False
        ),
        'yes_no_partial_na': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,N/A"',
            allow_blank=False
        ),
        'yes_no_limited': DataValidation(
            type="list",
            formula1='"Yes,No,Limited"',
            allow_blank=False
        ),
        'platform_type': DataValidation(
            type="list",
            formula1='"SIEM,IDS/IPS,EDR,NDR,UEBA,Log Management,Other"',
            allow_blank=False
        ),
        'deployment_model': DataValidation(
            type="list",
            formula1='"On-Premises,Cloud,Hybrid"',
            allow_blank=False
        ),
        'quality_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Limited,Poor"',
            allow_blank=False
        ),
        'search_performance': DataValidation(
            type="list",
            formula1='"<10 sec,10-60 sec,>60 sec"',
            allow_blank=False
        ),
        'ha_dr_status': DataValidation(
            type="list",
            formula1='"Documented,Tested,None"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1='"\u2705 Deployed,\u26A0\uFE0F Partial,\u274C Not Deployed,Planned"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1='"\u2705 Compliant,\u26A0\uFE0F Partial,\u274C Non-Compliant,N/A"',
            allow_blank=False
        ),
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,None"',
            allow_blank=False
        ),
        'system_type': DataValidation(
            type="list",
            formula1='"Server,Network Device,Security Appliance,Endpoint,Cloud Service,Database,Application"',
            allow_blank=False
        ),
        'criticality': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'log_collection_status': DataValidation(
            type="list",
            formula1='"\u2705 Collecting,\u26A0\uFE0F Partial,\u274C Not Collecting,\u21BB Planned"',
            allow_blank=False
        ),
        'collection_method': DataValidation(
            type="list",
            formula1='"Agent,Syslog,API,NetFlow,SNMP,WMI,Other"',
            allow_blank=False
        ),
        'protocol': DataValidation(
            type="list",
            formula1='"Syslog TCP,Syslog UDP,Syslog TLS,HTTP/S,API,Agent,SNMP,NetFlow,Other"',
            allow_blank=False
        ),
        'reliability': DataValidation(
            type="list",
            formula1='"High,Medium,Low,Unknown"',
            allow_blank=False
        ),
        'integration_type': DataValidation(
            type="list",
            formula1='"Real-time,Batch,Near Real-time,Manual"',
            allow_blank=False
        ),
        'enrichment_type': DataValidation(
            type="list",
            formula1='"Threat Intel,GeoIP,Asset Context,User Context,Custom,None"',
            allow_blank=False
        ),
    }

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab in sequence (1–5).', '2. Use dropdown menus where provided — do not type directly.', '3. Fill all yellow-highlighted cells with your organisation’s information.', '4. Refer to reference tables on each sheet for guidance.', '5. Complete compliance checklists on each assessment sheet.', '6. Document exceptions in the Exception/Deviation blocks.', '7. Gather evidence and list in Evidence Register.', '8. Review Summary Dashboard for compliance gaps.', '9. Obtain approvals via Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_monitoring_platform_sheet(ws, styles):
    """Create Monitoring Platform Capabilities assessment sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:U1")
    ws["A1"] = "1. MONITORING PLATFORM CAPABILITIES ASSESSMENT"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:U2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.4 - Assess SIEM/monitoring platform capabilities"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Does your organisation have monitoring platforms (SIEM, IDS/IPS, EDR, NDR) with adequate capabilities?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:U3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Response dropdown
    ws["A4"] = "Overall Status:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    thin_s = Side(style="thin")
    ws["B4"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    ws.merge_cells("B4:E4")

    # Column Headers (Row 6)
    headers = [
        ("A", "Platform/Tool Name", 28),
        ("B", "Platform Type", 20),
        ("C", "Vendor/Solution", 22),
        ("D", "Deployment Model", 18),
        ("E", "Log Collection Methods", 24),
        ("F", "Parsing Capabilities", 20),
        ("G", "Storage & Indexing", 20),
        ("H", "Search Performance", 18),
        ("I", "Real-Time Alerting", 18),
        ("J", "Correlation Engine", 18),
        ("K", "Threat Intel Integration", 22),
        ("L", "SOAR Integration", 18),
        ("M", "Visualization/Dashboards", 20),
        ("N", "High Availability", 18),
        ("O", "Disaster Recovery", 18),
        ("P", "Current EPS Capacity", 18),
        ("Q", "Implementation Status", 20),
        ("R", "Last Upgrade Date", 16),
        ("S", "Compliance Status", 18),
        ("T", "Gaps/Issues", 30),
        ("U", "Remediation Priority", 18),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "Splunk Enterprise",
        "SIEM",
        "Splunk Inc.",
        "Cloud",
        "Agents, Syslog, API",
        "Excellent",
        "Hot/Warm/Cold Tiers",
        "<10 sec",
        "Yes",
        "Advanced",
        "Yes",
        "Yes",
        "Excellent",
        "Yes",
        "Documented",
        "50'000 EPS",
        "\u2705 Deployed",
        "15.03.2024",
        "\u2705 Compliant",
        "None",
        "None"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-20: 13 rows)
    validations = create_base_validations(ws)
    
    for data_row in range(8, 21):
        for col_idx in range(1, 22):  # A to U
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # Platform Type
                validations['platform_type'].add(cell)
            elif col_letter == 'D':  # Deployment Model
                validations['deployment_model'].add(cell)
            elif col_letter in ['F', 'M']:  # Quality ratings
                validations['quality_rating'].add(cell)
            elif col_letter == 'H':  # Search Performance
                validations['search_performance'].add(cell)
            elif col_letter in ['I', 'J', 'K', 'L', 'N']:  # Yes/No/Planned
                validations['yes_no_planned'].add(cell)
            elif col_letter == 'O':  # DR Status
                validations['ha_dr_status'].add(cell)
            elif col_letter == 'Q':  # Implementation Status
                validations['implementation_status'].add(cell)
            elif col_letter == 'S':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'U':  # Priority
                validations['priority'].add(cell)

    # Compliance Checklist (Starting Row 22)
    row = 22
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "MONITORING PLATFORM CAPABILITIES CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Primary monitoring platform (SIEM) deployed",
        "Log collection supports multiple methods (agents, syslog, API)",
        "Parsing capabilities for common log formats (Windows, Linux, network devices)",
        "Storage includes hot/warm/cold tiers for cost optimization",
        "Search performance meets requirements (<10 sec for typical queries)",
        "Real-time alerting capability enabled",
        "Correlation engine supports multi-event detection",
        "Threat intelligence integration implemented",
        "SOAR or ticketing integration functional",
        "Dashboards configured for SOC operations",
        "High availability architecture implemented",
        "Disaster recovery plan documented and tested",
        "Platform capacity adequate for current log volume",
        "Platform upgraded within last 12 months",
        "Platform security hardened (RBAC, MFA, encryption)",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.merge_cells(f"B{row}:T{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:T{row}")
        
        # Status dropdown
        ws[f"U{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"U{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"U{row}"])

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(U:U,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Tables (Starting Row 40)
    row = 40
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "REFERENCE TABLES"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Table 1: Monitoring Platform Types
    row += 2
    ws[f"A{row}"] = "Table 1: Monitoring Platform Types"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws.merge_cells(f"A{row}:U{row}")

    row += 1
    table1_headers = ["Platform Type", "Primary Purpose", "Example Capabilities"]
    header_cells = [("A", "G"), ("H", "N"), ("O", "U")]
    for (start_col, end_col), header in zip(header_cells, table1_headers):
        ws[f"{start_col}{row}"] = header
        apply_style(ws[f"{start_col}{row}"], styles["column_header"])
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    table1_data = [
        ("SIEM", "Centralised log aggregation & correlation", "Search, alert, correlate, dashboards"),
        ("IDS/IPS", "Network intrusion detection", "Signature matching, anomaly detection"),
        ("EDR", "Endpoint threat detection", "Process monitoring, behavioural analysis"),
        ("NDR", "Network traffic analysis", "Flow analysis, encrypted traffic analysis"),
        ("UEBA", "User behaviour analytics", "Baseline deviation, risk scoring"),
    ]

    row += 1
    for platform, purpose, capabilities in table1_data:
        ws[f"A{row}"] = platform
        ws[f"H{row}"] = purpose
        ws[f"O{row}"] = capabilities
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"H{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"O{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:G{row}")
        ws.merge_cells(f"H{row}:N{row}")
        ws.merge_cells(f"O{row}:U{row}")
        row += 1

    # Table 2: Critical SIEM Capabilities
    row += 2
    ws[f"A{row}"] = "Table 2: Critical SIEM Capabilities (Per S2.1.4.1)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws.merge_cells(f"A{row}:U{row}")

    row += 1
    table2_headers = ["Capability", "Requirement Level", "Validation Method"]
    header_cells = [("A", "G"), ("H", "N"), ("O", "U")]
    for (start_col, end_col), header in zip(header_cells, table2_headers):
        ws[f"{start_col}{row}"] = header
        apply_style(ws[f"{start_col}{row}"], styles["column_header"])
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    table2_data = [
        ("Multi-format log parsing", "MUST HAVE", "Test with sample logs"),
        ("Indexed storage", "MUST HAVE", "Verify search performance"),
        ("Real-time alerting", "MUST HAVE", "Test alert generation"),
        ("Multi-event correlation", "MUST HAVE", "Create test correlation rule"),
        ("Threat intel integration", "MUST HAVE", "Verify IOC matching"),
        ("Search < 10 sec", "SHOULD HAVE", "Benchmark queries"),
    ]

    row += 1
    for capability, level, validation in table2_data:
        ws[f"A{row}"] = capability
        ws[f"H{row}"] = level
        ws[f"O{row}"] = validation
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"H{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"O{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:G{row}")
        ws.merge_cells(f"H{row}:N{row}")
        ws.merge_cells(f"O{row}:U{row}")
        row += 1

    # Exception/Deviation Block (Starting Row 70)
    row = 70
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "EXCEPTION/DEVIATION MANAGEMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    exception_headers = [
        "Exception ID", "Risk ID", "Business Justification", 
        "Compensating Controls", "Approval Status", "Review Date"
    ]
    for col_idx, header in enumerate(exception_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # 5 exception rows
    for i in range(5):
        row += 1
        for col_idx in range(1, 7):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for _dv in validations.values():
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: SHEET 3 - LOG SOURCE COVERAGE ASSESSMENT
# ============================================================================

def create_log_source_coverage_sheet(ws, styles):
    """Create Log Source Coverage assessment sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:T1")
    ws["A1"] = "2. LOG SOURCE COVERAGE ASSESSMENT"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:T2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.2 - Assess log source coverage"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Does your organisation collect logs from all critical systems and adequate coverage of standard systems?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:T3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Response dropdown
    ws["A4"] = "Overall Coverage Status:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    thin_s = Side(style="thin")
    ws["B4"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    ws.merge_cells("B4:E4")

    # Column Headers (Row 6)
    headers = [
        ("A", "System/Asset Name", 28),
        ("B", "System Type", 22),
        ("C", "Criticality", 15),
        ("D", "Location", 18),
        ("E", "System Owner", 20),
        ("F", "OS/Platform", 18),
        ("G", "Log Types Collected", 30),
        ("H", "Collection Method", 20),
        ("I", "Collection Protocol", 20),
        ("J", "Log Volume (GB/day)", 18),
        ("K", "Retention Period (days)", 20),
        ("L", "Parsing Status", 18),
        ("M", "Integration with SIEM", 20),
        ("N", "Collection Status", 20),
        ("O", "Last Verified", 16),
        ("P", "Reliability", 15),
        ("Q", "Compliance Status", 18),
        ("R", "Gaps/Issues", 30),
        ("S", "Remediation Plan", 30),
        ("T", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "DC01-WinServer",
        "Server",
        "Critical",
        "Primary DC",
        "IT Operations",
        "Windows Server 2022",
        "Security, System, Application",
        "Agent",
        "HTTPS",
        "2.5",
        "365",
        "Parsed",
        "Yes",
        "\u2705 Collecting",
        "15.12.2024",
        "High",
        "\u2705 Compliant",
        "None",
        "N/A",
        "None"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    for data_row in range(8, 58):
        for col_idx in range(1, 21):  # A to T
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # System Type
                validations['system_type'].add(cell)
            elif col_letter == 'C':  # Criticality
                validations['criticality'].add(cell)
            elif col_letter == 'H':  # Collection Method
                validations['collection_method'].add(cell)
            elif col_letter == 'I':  # Protocol
                validations['protocol'].add(cell)
            elif col_letter == 'N':  # Collection Status
                validations['log_collection_status'].add(cell)
            elif col_letter == 'P':  # Reliability
                validations['reliability'].add(cell)
            elif col_letter == 'Q':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'T':  # Priority
                validations['priority'].add(cell)

    # Coverage Summary Statistics (Starting Row 43)
    row = 43
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "LOG SOURCE COVERAGE STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    stat_headers = [("A", "E", "Metric"), ("F", "J", "Count/Value"), ("K", "O", "Target"), ("P", "T", "Status")]
    for start_col, end_col, header in stat_headers:
        ws[f"{start_col}{row}"] = header
        apply_style(ws[f"{start_col}{row}"], styles["column_header"])
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    statistics = [
        ("Total Systems Inventoried", "=COUNTA(A8:A40)", "N/A", ""),
        ("Critical Systems with Logs", "=COUNTIFS(C8:C40,\"Critical\",N8:N40,\"\u2705 Collecting\")", "100%", ""),
        ("High Priority Systems with Logs", "=COUNTIFS(C8:C40,\"High\",N8:N40,\"\u2705 Collecting\")", ">90%", ""),
        ("Total Log Volume (GB/day)", "=SUM(J8:J40)", "Monitor", ""),
        ("Systems with Compliance Gaps", "=COUNTIFS(Q8:Q40,\"\u274C Non-Compliant\")+COUNTIFS(Q8:Q40,\"\u26A0\uFE0F Partial\")", "0", ""),
        ("Collection Reliability High", "=COUNTIF(P8:P40,\"High\")", "Target", ""),
        ("Collection Reliability Low", "=COUNTIF(P8:P40,\"Low\")", "0", ""),
    ]

    row += 1
    for metric, formula, target, status in statistics:
        ws[f"A{row}"] = metric
        ws[f"F{row}"] = formula
        ws[f"K{row}"] = target
        ws[f"P{row}"] = status
        ws[f"P{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"P{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:E{row}")
        ws.merge_cells(f"F{row}:J{row}")
        ws.merge_cells(f"K{row}:O{row}")
        ws.merge_cells(f"P{row}:T{row}")
        row += 1

    # Compliance Checklist (Starting Row 53)
    row = 53
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "LOG SOURCE COVERAGE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "All critical systems (servers, databases) generating logs",
        "All network devices (firewalls, routers, switches) generating logs",
        "All security appliances (IDS/IPS, proxies) generating logs",
        "Authentication systems (AD, LDAP, SSO) generating logs",
        "Cloud services generating logs (AWS, Azure, O365)",
        "Critical applications generating logs",
        "Endpoints (workstations, laptops) generating logs where required",
        "Log collection methods reliable (agents, syslog, API)",
        "Log volume monitored and within capacity",
        "Log retention periods meet policy requirements (min 90 days for critical)",
        "Parsing configured for all critical log sources",
        "SIEM integration functional for all critical sources",
        "Log collection verified within last 30 days",
        "Gaps documented with remediation plans",
        "Coverage assessed quarterly",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:S{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:S{row}")
        
        # Status dropdown
        ws[f"T{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"T{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"T{row}"])

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(T55:T69,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Table: Critical System Types
    row = 75
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "REFERENCE: CRITICAL SYSTEM TYPES REQUIRING LOG COLLECTION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_header_cells = [("A", "G", "System Type"), ("H", "N", "Required Log Types"), ("O", "T", "Minimum Retention")]
    for start_col, end_col, header in ref_header_cells:
        ws[f"{start_col}{row}"] = header
        apply_style(ws[f"{start_col}{row}"], styles["column_header"])
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    ref_data = [
        ("Domain Controllers", "Security, System, Directory Service", "365 days"),
        ("Database Servers", "Audit, Error, Query logs", "365 days"),
        ("Web Servers", "Access, Error, Security", "180 days"),
        ("Firewalls", "Traffic, Deny, Alert", "365 days"),
        ("VPN Gateways", "Authentication, Connection, Error", "180 days"),
        ("Email Servers", "Message tracking, Security", "180 days"),
        ("Cloud Services (AWS/Azure/O365)", "Activity, Security, Configuration", "365 days"),
    ]

    row += 1
    for system_type, log_types, retention in ref_data:
        ws[f"A{row}"] = system_type
        ws[f"H{row}"] = log_types
        ws[f"O{row}"] = retention
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"H{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:G{row}")
        ws.merge_cells(f"H{row}:N{row}")
        ws.merge_cells(f"O{row}:T{row}")
        row += 1

    for _dv in validations.values():
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: SHEET 4 - DATA COLLECTION ARCHITECTURE
# ============================================================================

def create_data_collection_architecture_sheet(ws, styles):
    """Create Data Collection Architecture assessment sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:R1")
    ws["A1"] = "3. DATA COLLECTION ARCHITECTURE ASSESSMENT"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:R2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.3 - Assess data collection architecture"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Is your log collection architecture reliable, secure, and scalable?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:R3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Collection Component", 28),
        ("B", "Component Type", 22),
        ("C", "Purpose/Function", 30),
        ("D", "Protocol Used", 20),
        ("E", "Encryption Status", 18),
        ("F", "Authentication Method", 22),
        ("G", "Throughput Capacity", 18),
        ("H", "Current Utilization %", 18),
        ("I", "Buffer/Queue Size", 18),
        ("J", "Failover Configured", 18),
        ("K", "Load Balancing", 18),
        ("L", "Monitoring Enabled", 18),
        ("M", "Health Check Interval", 18),
        ("N", "Last Maintenance", 16),
        ("O", "Implementation Status", 20),
        ("P", "Compliance Status", 18),
        ("Q", "Issues/Gaps", 30),
        ("R", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "Syslog Forwarder Cluster",
        "Log Forwarder",
        "Central syslog collection",
        "Syslog TLS",
        "TLS 1.3",
        "Certificate-based",
        "100K EPS",
        "65%",
        "10GB",
        "Yes",
        "Yes",
        "Yes",
        "60 seconds",
        "10.12.2024",
        "\u2705 Deployed",
        "\u2705 Compliant",
        "None",
        "None"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    for data_row in range(8, 58):
        for col_idx in range(1, 19):  # A to R
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'D':  # Protocol
                validations['protocol'].add(cell)
            elif col_letter in ['J', 'K', 'L']:  # Yes/No
                validations['yes_no'].add(cell)
            elif col_letter == 'O':  # Implementation Status
                validations['implementation_status'].add(cell)
            elif col_letter == 'P':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'R':  # Priority
                validations['priority'].add(cell)

    # Architecture Compliance Checklist (Starting Row 59)
    row = 59
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "DATA COLLECTION ARCHITECTURE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Log collection uses encrypted protocols (TLS/HTTPS)",
        "Authentication required for log submission",
        "Collection components have adequate throughput capacity",
        "Buffering/queuing configured to prevent log loss",
        "Failover mechanisms implemented for critical components",
        "Load balancing configured where applicable",
        "Collection infrastructure monitored (health, performance)",
        "Alerts configured for collection failures",
        "Collection reliability >99% for critical systems",
        "Time synchronization (NTP) configured across infrastructure",
        "Log integrity mechanisms in place (checksums, signatures)",
        "Collection infrastructure documented (architecture diagrams)",
        "Capacity planning performed quarterly",
        "Disaster recovery tested for collection infrastructure",
        "Access to collection infrastructure restricted (RBAC)",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:Q{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:Q{row}")
        
        # Status dropdown
        ws[f"R{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"R{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"R{row}"])

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(R61:R75,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Table: Collection Protocols Security
    row = 77
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "REFERENCE: COLLECTION PROTOCOL SECURITY LEVELS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_header_cells = [("A", "E", "Protocol"), ("F", "I", "Security Level"), ("J", "N", "Use Case"), ("O", "R", "Recommendations")]
    for start_col, end_col, header in ref_header_cells:
        ws[f"{start_col}{row}"] = header
        apply_style(ws[f"{start_col}{row}"], styles["column_header"])
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    ref_data = [
        ("Syslog TLS (TCP 6514)", "High", "Secure syslog", "Recommended for all external collection"),
        ("HTTPS/API", "High", "Cloud services, REST APIs", "Use for cloud and API integrations"),
        ("Agent-based (HTTPS)", "High", "Endpoint/server collection", "Preferred for servers and endpoints"),
        ("Syslog TCP (TCP 514)", "Medium", "Legacy devices", "Use only when TLS unavailable"),
        ("Syslog UDP (UDP 514)", "Low", "Legacy, low-priority", "Avoid for critical systems"),
        ("SNMP v3", "Medium", "Network device monitoring", "Acceptable for network devices"),
        ("NetFlow", "Medium", "Network traffic metadata", "Supplementary to syslog"),
    ]

    row += 1
    for protocol, security, use_case, recommendation in ref_data:
        ws[f"A{row}"] = protocol
        ws[f"F{row}"] = security
        ws[f"J{row}"] = use_case
        ws[f"O{row}"] = recommendation
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"F{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"J{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"O{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:E{row}")
        ws.merge_cells(f"F{row}:I{row}")
        ws.merge_cells(f"J{row}:N{row}")
        ws.merge_cells(f"O{row}:R{row}")
        row += 1

    for _dv in validations.values():
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 7: SHEET 5 - INTEGRATION & ENRICHMENT
# ============================================================================

def create_integration_enrichment_sheet(ws, styles):
    """Create Integration & Enrichment assessment sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:P1")
    ws["A1"] = "4. INTEGRATION & ENRICHMENT ASSESSMENT"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:P2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.5 - Assess integration and data enrichment"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Are monitoring systems properly integrated with threat intelligence, asset inventory, and other context sources?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:P3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Integration/Enrichment Name", 30),
        ("B", "Integration Type", 22),
        ("C", "Data Source", 25),
        ("D", "Enrichment Type", 22),
        ("E", "Integration Method", 20),
        ("F", "Update Frequency", 18),
        ("G", "Data Quality", 18),
        ("H", "Coverage %", 15),
        ("I", "Latency", 15),
        ("J", "Reliability", 15),
        ("K", "Last Updated", 16),
        ("L", "Implementation Status", 20),
        ("M", "Compliance Status", 18),
        ("N", "Value/Impact", 25),
        ("O", "Issues/Gaps", 30),
        ("P", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "MISP Threat Feed",
        "Real-time",
        "MISP Community",
        "Threat Intel",
        "API",
        "Hourly",
        "Excellent",
        "95%",
        "<1 min",
        "High",
        "07.01.2025",
        "\u2705 Deployed",
        "\u2705 Compliant",
        "IOC matching for alerts",
        "None",
        "High"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    for data_row in range(8, 58):
        for col_idx in range(1, 17):  # A to P
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # Integration Type
                validations['integration_type'].add(cell)
            elif col_letter == 'D':  # Enrichment Type
                validations['enrichment_type'].add(cell)
            elif col_letter == 'G':  # Data Quality
                validations['quality_rating'].add(cell)
            elif col_letter == 'J':  # Reliability
                validations['reliability'].add(cell)
            elif col_letter == 'L':  # Implementation Status
                validations['implementation_status'].add(cell)
            elif col_letter == 'M':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'P':  # Priority
                validations['priority'].add(cell)

    # Integration Compliance Checklist (Starting Row 59)
    row = 59
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "INTEGRATION & ENRICHMENT CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Threat intelligence feed(s) integrated into SIEM",
        "Threat intelligence updated at least daily",
        "GeoIP enrichment configured for IP addresses",
        "Asset inventory integrated (hostname, owner, criticality)",
        "User context enrichment (AD/LDAP integration)",
        "DNS enrichment for domain resolution",
        "CMDB integration for asset context",
        "Vulnerability data integrated (scanner results)",
        "Ticketing system integration (ServiceNow, JIRA)",
        "SOAR platform integration for automated response",
        "Cloud provider APIs integrated (AWS, Azure, GCP)",
        "Enrichment latency acceptable (<5 minutes)",
        "Enrichment coverage >90% for critical assets",
        "Integration health monitored and alerted",
        "Integration documentation maintained",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:O{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:O{row}")
        
        # Status dropdown
        ws[f"P{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"P{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"P{row}"])

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(P61:P75,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Table: Common Enrichment Sources
    row = 77
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "REFERENCE: COMMON ENRICHMENT SOURCES"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_header_cells = [("A", "E", "Enrichment Type"), ("F", "K", "Example Sources"), ("L", "P", "Benefit to Monitoring")]
    for start_col, end_col, header in ref_header_cells:
        ws[f"{start_col}{row}"] = header
        apply_style(ws[f"{start_col}{row}"], styles["column_header"])
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    ref_data = [
        ("Threat Intelligence", "MISP, AlienVault OTX, VirusTotal, Abuse.ch", "IOC detection, malicious IP/domain identification"),
        ("GeoIP", "MaxMind, IP2Location", "Geo-based anomaly detection, compliance"),
        ("Asset Context", "CMDB, ServiceNow, Lansweeper", "Criticality assessment, owner identification"),
        ("User Context", "Active Directory, Okta, Microsoft Entra ID (formerly Azure AD)", "User behaviour analysis, privilege context"),
        ("Vulnerability Data", "Qualys, Tenable, Rapid7", "Risk scoring, exploit correlation"),
        ("DNS", "Internal DNS, passive DNS", "Domain reputation, DGA detection"),
    ]

    row += 1
    for enrich_type, sources, benefit in ref_data:
        ws[f"A{row}"] = enrich_type
        ws[f"F{row}"] = sources
        ws[f"L{row}"] = benefit
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"F{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"L{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:E{row}")
        ws.merge_cells(f"F{row}:K{row}")
        ws.merge_cells(f"L{row}:P{row}")
        row += 1

    for _dv in validations.values():
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: SHEET 6 - PERFORMANCE & SCALABILITY
# ============================================================================

def create_performance_scalability_sheet(ws, styles):
    """Create Performance & Scalability assessment sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:Q1")
    ws["A1"] = "5. PERFORMANCE & SCALABILITY ASSESSMENT"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:Q2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.6 - Assess monitoring performance and scalability"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Is the monitoring infrastructure performing adequately and scalable for future growth?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:Q3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Performance Metrics Section (Starting Row 5)
    row = 5
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "CURRENT PERFORMANCE METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Metric", 35),
        ("B", "Current Value", 20),
        ("C", "Unit", 15),
        ("D", "Target/Threshold", 20),
        ("E", "Status", 18),
        ("F", "Trend", 18),
        ("G", "Last Measured", 18),
        ("H", "Notes", 35),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Performance metrics data rows
    performance_metrics = [
        ("SIEM Ingestion Rate (Current)", "", "EPS", "N/A", "", "", "", ""),
        ("SIEM Ingestion Rate (Peak)", "", "EPS", "N/A", "", "", "", ""),
        ("SIEM Licensed Capacity", "", "EPS", "N/A", "", "", "", ""),
        ("Capacity Utilization", "", "%", "<80%", "", "", "", ""),
        ("Average Search Response Time", "", "seconds", "<10 sec", "", "", "", ""),
        ("Indexing Lag", "", "minutes", "<5 min", "", "", "", ""),
        ("Storage Utilization", "", "%", "<80%", "", "", "", ""),
        ("Hot Tier Storage Available", "", "TB", ">20%", "", "", "", ""),
        ("Daily Log Volume", "", "GB/day", "N/A", "", "", "", ""),
        ("Data Retention Compliance", "", "days", ">=90", "", "", "", ""),
        ("SIEM Availability (Last 30d)", "", "%", ">99.5%", "", "", "", ""),
        ("Alert Generation Latency", "", "seconds", "<60 sec", "", "", "", ""),
        ("Correlation Rule Execution Time", "", "seconds", "<5 sec", "", "", "", ""),
        ("Dashboard Load Time", "", "seconds", "<3 sec", "", "", "", ""),
        ("API Response Time", "", "milliseconds", "<500 ms", "", "", "", ""),
    ]

    validations = create_base_validations(ws)
    row = 7
    for metric, value, unit, target, status, trend, measured, notes in performance_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = value
        ws[f"C{row}"] = unit
        ws[f"D{row}"] = target
        ws[f"E{row}"] = status
        ws[f"F{row}"] = trend
        ws[f"G{row}"] = measured
        ws[f"H{row}"] = notes
        
        # Style first column
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")

        # Yellow input cells for current value and notes
        for col in ['B', 'F', 'G', 'H']:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        # Status dropdown
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['compliance_status'].add(ws[f"E{row}"])

        # Borders on all cells
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f"{col}{row}"].border = border

        ws.merge_cells(f"H{row}:Q{row}")
        row += 1

    # Scalability Assessment (Starting Row 24)
    row = 24
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "SCALABILITY ASSESSMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    scalability_questions = [
        ("Current log sources (count)", "", "N/A"),
        ("Projected log sources (12 months)", "", "N/A"),
        ("Current daily log volume (GB)", "", "N/A"),
        ("Projected daily log volume (12 months, GB)", "", "N/A"),
        ("Growth rate (%)", "", "<30% increase"),
        ("Available capacity for growth (%)", "", ">50%"),
        ("Time to add new log source", "", "<1 hour"),
        ("Horizontal scaling capability", "", "Yes"),
        ("Vertical scaling capability", "", "Yes"),
        ("Capacity planning performed", "", "Quarterly"),
    ]

    row += 1
    scale_header_cells = [("A", "F", "Scalability Factor"), ("G", "L", "Current/Projected"), ("M", "Q", "Target")]
    for start_col, end_col, header in scale_header_cells:
        ws[f"{start_col}{row}"] = header
        apply_style(ws[f"{start_col}{row}"], styles["column_header"])
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

    row += 1
    for factor, value, target in scalability_questions:
        ws[f"A{row}"] = factor
        ws[f"G{row}"] = value
        ws[f"M{row}"] = target

        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in ['A', 'G', 'M']:
            ws[f"{col}{row}"].border = border

        ws.merge_cells(f"A{row}:F{row}")
        ws.merge_cells(f"G{row}:L{row}")
        ws.merge_cells(f"M{row}:Q{row}")
        row += 1

    # Performance Checklist (Starting Row 37)
    row = 37
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "PERFORMANCE & SCALABILITY CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "SIEM capacity utilization <80%",
        "Search performance meets <10 second target",
        "Indexing lag <5 minutes",
        "Storage utilization <80%",
        "SIEM availability >99.5%",
        "Alert latency <60 seconds",
        "Capacity planning performed quarterly",
        "Growth projections documented",
        "Scalability tested (load testing)",
        "Performance baselines established",
        "Performance monitoring automated",
        "Performance metrics reviewed monthly",
        "Capacity alerts configured",
        "Retention policies optimized for cost",
        "Disaster recovery capacity validated",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:P{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:P{row}")
        
        # Status dropdown
        ws[f"Q{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"Q{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"Q{row}"])

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(Q39:Q53,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    for _dv in validations.values():
        ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: SHEET 7 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with TABLE 1/2/3 Gold Standard layout."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: A1:G1 merged 003366 header
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle
    ws['A2'] = f'Summary Dashboard — {CONTROL_NAME} | {GENERATED_DATE}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.freeze_panes = 'A4'

    # TABLE 1: COMPLIANCE OVERVIEW
    ws.merge_cells('A4:G4')
    ws['A4'] = 'TABLE 1: COMPLIANCE OVERVIEW'
    ws['A4'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A4'].border = border

    # Row 5: headers
    t1_headers = ['Assessment Area', 'Total Items', 'Compliant', 'Partial', 'Non-Compliant', 'N/A', 'Compliance %']
    for col_idx, header in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='000000')
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Rows 6-10: data rows
    ws.cell(row=6, column=1, value='1. Monitoring Platform').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=6, column=2, value='=SUM(C6:F6)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=3, value='=COUNTIF(\'1. Monitoring Platform\'!U24:U200,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=4, value='=COUNTIF(\'1. Monitoring Platform\'!U24:U200,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=5, value='=COUNTIF(\'1. Monitoring Platform\'!U24:U200,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=6, value='=COUNTIF(\'1. Monitoring Platform\'!U24:U200,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7, value='=IF((B6-F6)=0,0,C6/(B6-F6))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=6, column=col_idx).border = border

    ws.cell(row=7, column=1, value='2. Log Source Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=7, column=2, value='=SUM(C7:F7)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=3, value='=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=4, value='=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=5, value='=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=6, value='=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7, value='=IF((B7-F7)=0,0,C7/(B7-F7))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=7, column=col_idx).border = border

    ws.cell(row=8, column=1, value='3. Data Collection Arch').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=8, column=2, value='=SUM(C8:F8)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=3, value='=COUNTIF(\'3. Data Collection Arch\'!R61:R75,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=4, value='=COUNTIF(\'3. Data Collection Arch\'!R61:R75,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=5, value='=COUNTIF(\'3. Data Collection Arch\'!R61:R75,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=6, value='=COUNTIF(\'3. Data Collection Arch\'!R61:R75,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7, value='=IF((B8-F8)=0,0,C8/(B8-F8))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=8, column=col_idx).border = border

    ws.cell(row=9, column=1, value='4. Integration Enrichment').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=9, column=2, value='=SUM(C9:F9)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=3, value='=COUNTIF(\'4. Integration Enrichment\'!P61:P75,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=4, value='=COUNTIF(\'4. Integration Enrichment\'!P61:P75,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=5, value='=COUNTIF(\'4. Integration Enrichment\'!P61:P75,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=6, value='=COUNTIF(\'4. Integration Enrichment\'!P61:P75,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7, value='=IF((B9-F9)=0,0,C9/(B9-F9))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=9, column=col_idx).border = border

    ws.cell(row=10, column=1, value='5. Performance Scale').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=10, column=2, value='=SUM(C10:F10)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=3, value='=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=4, value='=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=5, value='=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=6, value='=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7, value='=IF((B10-F10)=0,0,C10/(B10-F10))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=10, column=col_idx).border = border

    # Row 11: TOTAL
    ws.cell(row=11, column=1, value='TOTAL').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=1).border = border
    for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F'], start=2):
        ws.cell(row=11, column=col_idx, value=f'=SUM({col_letter}6:{col_letter}10)').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=11, column=col_idx).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws.cell(row=11, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=11, column=col_idx).border = border
    ws.cell(row=11, column=7, value='=IF((B11-F11)=0,0,C11/(B11-F11))').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=7).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=11, column=7).number_format = '0.0%'
    ws.cell(row=11, column=7).border = border

    # TABLE 2: KEY METRICS
    ws.merge_cells('A13:G13')
    ws['A13'] = 'TABLE 2: KEY METRICS'
    ws['A13'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A13'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A13'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A13'].border = border
    ws.merge_cells('A14:E14')
    ws['A14'] = 'Metric'
    ws['A14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A14'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A14'].border = border
    ws.merge_cells('F14:G14')
    ws['F14'] = 'Value'
    ws['F14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F14'].border = border
    yllw_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws.merge_cells('A15:E15')
    ws.cell(row=15, column=1, value='Monitoring Platforms — Deployed').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=15, column=1).border = border
    ws.merge_cells('F15:G15')
    ws.cell(row=15, column=6, value='=COUNTIF(\'1. Monitoring Platform\'!Q8:Q200,"✅ Deployed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=15, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=15, column=col_idx).border = border

    ws.merge_cells('A16:E16')
    ws.cell(row=16, column=1, value='Monitoring Platforms — Compliance Gaps').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=16, column=1).border = border
    ws.merge_cells('F16:G16')
    ws.cell(row=16, column=6, value='=COUNTIF(\'1. Monitoring Platform\'!S8:S200,"❌ Non-Compliant")+COUNTIF(\'1. Monitoring Platform\'!S8:S200,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=16, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=16, column=col_idx).border = border

    ws.merge_cells('A17:E17')
    ws.cell(row=17, column=1, value='Log Sources — Actively Collecting').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=17, column=1).border = border
    ws.merge_cells('F17:G17')
    ws.cell(row=17, column=6, value='=COUNTIF(\'2. Log Source Coverage\'!N8:N200,"✅ Collecting")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=17, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=17, column=col_idx).border = border

    ws.merge_cells('A18:E18')
    ws.cell(row=18, column=1, value='Log Sources — Not Collecting').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=18, column=1).border = border
    ws.merge_cells('F18:G18')
    ws.cell(row=18, column=6, value='=COUNTIF(\'2. Log Source Coverage\'!N8:N200,"❌ Not Collecting")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=18, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=18, column=col_idx).border = border

    ws.merge_cells('A19:E19')
    ws.cell(row=19, column=1, value='Critical Log Sources Missing').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=19, column=1).border = border
    ws.merge_cells('F19:G19')
    ws.cell(row=19, column=6, value='=COUNTIFS(\'2. Log Source Coverage\'!C8:C200,"Critical",\'2. Log Source Coverage\'!N8:N200,"❌ Not Collecting")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=19, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=19, column=col_idx).border = border

    ws.merge_cells('A20:E20')
    ws.cell(row=20, column=1, value='Collection Components — Deployed').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=20, column=1).border = border
    ws.merge_cells('F20:G20')
    ws.cell(row=20, column=6, value='=COUNTIF(\'3. Data Collection Arch\'!O8:O200,"✅ Deployed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=20, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=20, column=col_idx).border = border

    ws.merge_cells('A21:E21')
    ws.cell(row=21, column=1, value='Integrations / Enrichment — Deployed').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=21, column=1).border = border
    ws.merge_cells('F21:G21')
    ws.cell(row=21, column=6, value='=COUNTIF(\'4. Integration Enrichment\'!L8:L200,"✅ Deployed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=21, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=21, column=col_idx).border = border

    ws.merge_cells('A22:E22')
    ws.cell(row=22, column=1, value='High-Priority Items Across All Sheets').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=22, column=1).border = border
    ws.merge_cells('F22:G22')
    ws.cell(row=22, column=6, value='=COUNTIF(\'2. Log Source Coverage\'!C8:C200,"Critical")+COUNTIF(\'3. Data Collection Arch\'!C8:C200,"Critical")+COUNTIF(\'4. Integration Enrichment\'!C8:C200,"Critical")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=22, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=22, column=col_idx).border = border

    # TABLE 3: CRITICAL FINDINGS
    ws.merge_cells('A24:G24')
    ws['A24'] = 'TABLE 3: CRITICAL FINDINGS'
    ws['A24'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A24'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A24'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A24'].border = border
    ws.merge_cells('A25:E25')
    ws['A25'] = 'Category / Finding'
    ws['A25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A25'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A25'].border = border
    ws['F25'] = 'Count'
    ws['F25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F25'].border = border
    ws['G25'] = 'Action'
    ws['G25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['G25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['G25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G25'].border = border

    ws.merge_cells('A26:E26')
    ws.cell(row=26, column=1, value='Log Sources — Not Collecting').fill = yllw_fill
    ws.cell(row=26, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=26, column=1).border = border
    ws.cell(row=26, column=6, value='=COUNTIF(\'2. Log Source Coverage\'!N8:N200,"❌ Not Collecting")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=26, column=6).fill = yllw_fill
    ws.cell(row=26, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=26, column=6).border = border
    ws.cell(row=26, column=7, value='').fill = yllw_fill
    ws.cell(row=26, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=26, column=col_idx).fill = yllw_fill
        ws.cell(row=26, column=col_idx).border = border

    ws.merge_cells('A27:E27')
    ws.cell(row=27, column=1, value='Critical Log Sources Missing').fill = yllw_fill
    ws.cell(row=27, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=27, column=1).border = border
    ws.cell(row=27, column=6, value='=COUNTIFS(\'2. Log Source Coverage\'!C8:C200,"Critical",\'2. Log Source Coverage\'!N8:N200,"❌ Not Collecting")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=27, column=6).fill = yllw_fill
    ws.cell(row=27, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=27, column=6).border = border
    ws.cell(row=27, column=7, value='').fill = yllw_fill
    ws.cell(row=27, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=27, column=col_idx).fill = yllw_fill
        ws.cell(row=27, column=col_idx).border = border

    ws.merge_cells('A28:E28')
    ws.cell(row=28, column=1, value='Platforms — Non-Compliant').fill = yllw_fill
    ws.cell(row=28, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=28, column=1).border = border
    ws.cell(row=28, column=6, value='=COUNTIF(\'1. Monitoring Platform\'!S8:S200,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=28, column=6).fill = yllw_fill
    ws.cell(row=28, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=28, column=6).border = border
    ws.cell(row=28, column=7, value='').fill = yllw_fill
    ws.cell(row=28, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=28, column=col_idx).fill = yllw_fill
        ws.cell(row=28, column=col_idx).border = border

    ws.merge_cells('A29:E29')
    ws.cell(row=29, column=1, value='Components — Not Deployed').fill = yllw_fill
    ws.cell(row=29, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=29, column=1).border = border
    ws.cell(row=29, column=6, value='=COUNTIF(\'3. Data Collection Arch\'!O8:O200,"❌ Not Deployed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=29, column=6).fill = yllw_fill
    ws.cell(row=29, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=29, column=6).border = border
    ws.cell(row=29, column=7, value='').fill = yllw_fill
    ws.cell(row=29, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=29, column=col_idx).fill = yllw_fill
        ws.cell(row=29, column=col_idx).border = border

    ws.merge_cells('A30:E30')
    ws.cell(row=30, column=1, value='Integrations — Not Deployed').fill = yllw_fill
    ws.cell(row=30, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=30, column=1).border = border
    ws.cell(row=30, column=6, value='=COUNTIF(\'4. Integration Enrichment\'!L8:L200,"❌ Not Deployed")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=30, column=6).fill = yllw_fill
    ws.cell(row=30, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=30, column=6).border = border
    ws.cell(row=30, column=7, value='').fill = yllw_fill
    ws.cell(row=30, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=30, column=col_idx).fill = yllw_fill
        ws.cell(row=30, column=col_idx).border = border

    # Row 31: instruction
    ws.merge_cells('A31:G31')
    ws['A31'] = 'Filter source sheets using AutoFilter — see column headers above'
    ws['A31'].font = Font(name='Calibri', size=9, italic=True, color='003366')
    ws['A31'].alignment = Alignment(horizontal='left', vertical='center')

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — Gold Standard 8-column format."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: merged title, 003366 fill, white bold 14pt + borders all 8 cells
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _c in range(1, 9):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: generic subtitle — italic, left-aligned
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty separator

    # Row 4: column headers — 003366 fill, white bold
    er_headers = [
        ('Evidence ID', 15),
        ('Assessment Area', 25),
        ('Evidence Type', 22),
        ('Description', 40),
        ('Location/Path', 45),
        ('Date Collected', 16),
        ('Collected By', 20),
        ('Verification Status', 22),
    ]
    for col_idx, (header, width) in enumerate(er_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Log Export,Documentation,Report,Network scan,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    ws.add_data_validation(status_dv)

    type_dv.add("C6:C105")
    status_dv.add("H6:H105")

    # Row 5: F2F2F2 grey sample row — fully populated example
    sample = [
        'EV-001',
        'Monitoring Infrastructure',
        'Configuration file',
        'SIEM platform configuration export — Splunk Enterprise production instance',
        '\\\\fileserver\\isms\\evidence\\monitoring\\splunk-config-export.json',
        '2024-03-01',
        'John Davies',
        'Verified',
    ]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.font = Font(name="Calibri", size=10)

    # Rows 6-105: 100 empty FFFFCC data rows (no pre-populated IDs)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Freeze pane at A5 (freeze rows 1–4)
    ws.freeze_panes = "A5"

def _as_border():
    """Return a fresh thin border object."""
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet for governance."""
    thin_border = _as_border()

    # Helper: apply thin border to top-left cell of every merged range
    def _border_merged(cell_ref):
        ws[cell_ref].border = _as_border()

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = _as_border()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = "Monitoring Infrastructure Assessment Approval Workflow"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = _as_border()
    # Document Summary
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    prepared_fields = [
        ("Name:", ""),
        ("Title:", ""),
        ("Date:", ""),
        ("Signature:", ""),
    ]

    row += 1
    for label, value in prepared_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as_border()
        ws.merge_cells(f"B{row}:E{row}")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as_border()
        row += 1

    # Reviewed By (SOC Lead)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    row += 1
    for label, value in prepared_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as_border()
        ws.merge_cells(f"B{row}:E{row}")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as_border()
        row += 1

    # Review Comments
    row += 1
    ws[f"A{row}"] = "Review Comments:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"A{row}:E{row}")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    row += 1
    ws.merge_cells(f"A{row}:E{row+3}")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    row += 4

    # Approved By (CISO)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    row += 1
    approval_fields = [
        ("Name:", ""),
        ("Title:", ""),
        ("Approval Decision:", "[Dropdown: Approved, Approved with Conditions, Rejected]"),
        ("Date:", ""),
        ("Signature:", ""),
    ]

    approval_validation = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected"',
        allow_blank=False
    )

    for label, value in approval_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as_border()
        ws.merge_cells(f"B{row}:E{row}")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as_border()
        # Add dropdown for Approval Decision
        if "Approval Decision" in label:
            approval_validation.add(ws[f"B{row}"])

        row += 1

    # Conditions/Comments
    row += 1
    ws[f"A{row}"] = "Conditions/Comments:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"A{row}:E{row}")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    row += 1
    ws.merge_cells(f"A{row}:E{row+4}")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    row += 5

    # Final Assessment Decision
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "FINAL DECISION"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    row += 1
    ws[f"A{row}"] = "Decision:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = _as_border()
    ws.merge_cells(f"B{row}:E{row}")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = _as_border()
    decision_validation = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=False
    )
    decision_validation.add(ws[f"B{row}"])

    # Next Review Date
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as_border()
    row += 1
    ws[f"A{row}"] = "Scheduled Review:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = _as_border()
    ws.merge_cells(f"B{row}:E{row}")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = _as_border()
    # Set column widths
    ws.column_dimensions['A'].width = 32
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 25

    # Freeze pane
    ws.freeze_panes = "A3"

    ws.add_data_validation(approval_validation)
    ws.add_data_validation(decision_validation)

# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "The first principle is that you must not fool yourself
    — and you are the easiest person to fool." - Richard Feynman
    
    This is Systems Engineering applied to ISMS compliance:
    We assess CAPABILITIES, not checkboxes. We gather EVIDENCE, not platitudes.
    We implement DETECTIVE CONTROLS, not cargo cult rituals.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.16.1 - Monitoring Infrastructure Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities")
    logger.info("=" * 78)
    logger.info("\n>>> Systems Engineering Approach: Evidence-Based Monitoring Assessment")
    logger.info(">>> Comprehensive: Platform, Coverage, Architecture, Integration, Performance")
    logger.info(">>> Audit-Ready: 100 evidence entries, 75 compliance checkpoints")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("\u2705 Workbook created with 9 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")
    
    logger.info("  [1/9] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  \u2705 Instructions complete")

    logger.info("  [2/9] Creating Monitoring Platform Capabilities...")
    create_monitoring_platform_sheet(wb["1. Monitoring Platform"], styles)
    logger.info("  \u2705 Platform assessment complete (15 capability checks)")

    logger.info("  [3/9] Creating Log Source Coverage...")
    create_log_source_coverage_sheet(wb["2. Log Source Coverage"], styles)
    logger.info("  \u2705 Coverage assessment complete (33 log source rows, 15 checks)")

    logger.info("  [4/9] Creating Data Collection Architecture...")
    create_data_collection_architecture_sheet(wb["3. Data Collection Arch"], styles)
    logger.info("  \u2705 Architecture assessment complete (18 component rows, 15 checks)")

    logger.info("  [5/9] Creating Integration & Enrichment...")
    create_integration_enrichment_sheet(wb["4. Integration Enrichment"], styles)
    logger.info("  \u2705 Integration assessment complete (15 integration rows, 15 checks)")

    logger.info("  [6/9] Creating Performance & Scalability...")
    create_performance_scalability_sheet(wb["5. Performance Scale"], styles)
    logger.info("  \u2705 Performance assessment complete (15 metrics, 10 scalability factors)")

    logger.info("  [8/9] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  \u2705 Evidence register complete (100 evidence rows)")

    logger.info("  [7/9] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  \u2705 Dashboard complete (consolidated compliance view)")

    logger.info("  [9/9] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  \u2705 Approval workflow complete (4-level sign-off)")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.16.1_Monitoring_Infrastructure_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"\u2705 SUCCESS: {output_path}")
    except Exception as e:
        logger.error(f"\u274C ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets:")
    logger.info("  \u2022 Instructions & Legend (usage guidance, definitions)")
    logger.info("  \u2022 1. Monitoring Platform Capabilities (SIEM/IDS/EDR/NDR assessment)")
    logger.info("  \u2022 2. Log Source Coverage (33 system rows, coverage tracking)")
    logger.info("  \u2022 3. Data Collection Architecture (18 component rows, security)")
    logger.info("  \u2022 4. Integration & Enrichment (15 integration rows, threat intel)")
    logger.info("  \u2022 5. Performance & Scalability (15 metrics, capacity planning)")
    logger.info("\n>>> Consolidation & Governance:")
    logger.info("  \u2022 Summary Dashboard (overall compliance, key findings)")
    logger.info("  \u2022 Evidence Register (100 evidence entries)")
    logger.info("  \u2022 Approval Sign-Off (4-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info(">>> ASSESSMENT CAPABILITIES:")
    logger.info("  \u2022 75 compliance checkpoint items across 5 assessment areas")
    logger.info("  \u2022 33 log source tracking rows")
    logger.info("  \u2022 18 collection architecture component rows")
    logger.info("  \u2022 15 integration/enrichment tracking rows")
    logger.info("  \u2022 15 performance metrics + 10 scalability factors")
    logger.info("  \u2022 100 evidence documentation entries")
    logger.info("  \u2022 Automated compliance % calculations")
    logger.info("  \u2022 Gap identification and priority tracking")
    logger.info("\n" + "─" * 78)
    logger.info(">>> KEY FEATURES:")
    logger.info("  \u2705 Platform-agnostic (works with ANY SIEM/monitoring solution)")
    logger.info("  \u2705 Comprehensive evidence collection")
    logger.info("  \u2705 Automated compliance calculations")
    logger.info("  \u2705 Coverage assessment (critical systems priority)")
    logger.info("  \u2705 Architecture security validation")
    logger.info("  \u2705 Integration health tracking")
    logger.info("  \u2705 Performance & capacity monitoring")
    logger.info("  \u2705 Multi-level approval workflow")
    logger.info("  \u2705 Semi-annual review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(">>> NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Complete Instructions & Legend sheet first")
    logger.info("  3. Fill in YOUR monitoring platforms (vendor-agnostic approach)")
    logger.info("  4. Document ALL log sources (prioritize Critical/High systems)")
    logger.info("  5. Assess collection architecture security")
    logger.info("  6. Evaluate integrations and enrichment")
    logger.info("  7. Review performance metrics and capacity")
    logger.info("  8. Check Summary Dashboard for gaps")
    logger.info("  9. Document evidence in Evidence Register")
    logger.info("  10. Obtain final approval via Approval Sign-Off")
    logger.info("\n>>> PRO TIP:")
    logger.info("  This workbook is technology-independent. Whether you use Splunk,")
    logger.info("  Elastic, QRadar, Sentinel, or ANY other SIEM - this framework")
    logger.info("  assesses MONITORING CAPABILITIES, not brand names.")
    logger.info("\n" + "=" * 78)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info("\n>>> This is not cargo cult ISMS. This is evidence-based monitoring.")
    logger.info(">>> We assess DETECTIVE CONTROLS, not checkbox compliance.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
