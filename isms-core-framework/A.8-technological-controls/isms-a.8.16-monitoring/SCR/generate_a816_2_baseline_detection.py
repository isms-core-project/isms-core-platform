#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.16.2 - Baseline & Anomaly Detection Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities
Assessment Domain 2 of 5: Baseline Definition and Anomaly Detection Capabilities

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific baseline definitions, detection rules, and
threat intelligence requirements.

Key customisation areas:
1. Baseline profile definitions (match your normal behavior patterns)
2. Detection rule catalog (specific to your threat landscape)
3. ML/AI model configurations (aligned with your SIEM capabilities)
4. Threat intelligence feeds (per your subscriptions)
5. Validation testing criteria (based on your security operations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for monitoring)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
baseline definition and anomaly detection capabilities against ISO 27001:2022 
Control A.8.16 requirements.

**Purpose:**
Enables systematic assessment of normal behavior baselines, detection rule
effectiveness, ML/AI model capabilities, threat intelligence integration, and
validation testing against Control A.8.16 detection requirements.

**Assessment Scope:**
- Normal behavior profile definition and maintenance
- Detection rule effectiveness and coverage
- Machine learning and AI model assessment
- Threat intelligence integration
- Validation testing and false positive management
- Gap analysis and detection tuning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and detection standards
2. Normal Behavior Profiles - Baseline definition and maintenance
3. Detection Rules - Rule effectiveness and coverage assessment
4. ML/AI Models - Machine learning detection capabilities
5. Threat Intelligence - Threat intel integration and utilization
6. Validation Testing - Detection effectiveness testing results
7. Summary Dashboard - Detection maturity and effectiveness metrics
8. Evidence Register - Audit evidence tracking and documentation
9. Approval Sign-Off - Multi-level approval workflow

**Key Features:**
- Data validation with detection methodology dropdown lists
- Conditional formatting for detection effectiveness visualization
- Automated false positive rate calculations
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with A.8.16 Compliance Dashboard

**Integration:**
This assessment feeds into ISMS-IMP-A.8.16.5 Compliance Dashboard, which
consolidates data from all five monitoring assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a816_2_baseline_detection.py

Output:
    File: ISMS_A_8_16_2_Baseline_Detection_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Document normal behavior baselines (systems, users, networks)
    2. Inventory detection rules and effectiveness metrics
    3. Assess ML/AI model capabilities and accuracy
    4. Review threat intelligence integration
    5. Analyze validation testing results
    6. Calculate false positive rates and detection effectiveness
    7. Conduct gap analysis for detection blind spots
    8. Define tuning and remediation actions
    9. Collect and link audit evidence
    10. Obtain stakeholder approvals
    11. Feed results into A.8.16.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.16
Assessment Domain:    2 of 5 (Baseline & Anomaly Detection)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.16: Monitoring Activities Policy (Governance)
    - ISMS-POL-A.8.16-S2.2: Baseline & Anomaly Detection Requirements
    - ISMS-IMP-A.8.16.2: Baseline & Detection Implementation Guide
    - ISMS-IMP-A.8.16.1: Monitoring Infrastructure Assessment (Domain 1)
    - ISMS-IMP-A.8.16.3: Coverage Assessment (Domain 3)
    - ISMS-IMP-A.8.16.4: Alert Management Assessment (Domain 4)
    - ISMS-IMP-A.8.16.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.16.2 specification
    - Supports comprehensive baseline and detection evaluation
    - Integrated with A.8.16.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Detection Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This is not cargo cult ISMS. Detection without baselines is noise generation.
If you don't know normal, you can't detect abnormal. Validate everything.

**False Positive Management:**
High false positive rates destroy SOC effectiveness. Target <5% false positive
rate for production detection rules. Untuned detection is often worse than
no detection - it creates alert fatigue and masks real threats.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will expect:
- Documented baselines with update procedures
- Detection effectiveness metrics (true positive rate, false positive rate)
- Validation testing results
- Continuous improvement evidence

**Data Protection:**
Assessment workbooks contain sensitive security details including:
- Detection methodologies and rules
- Threat intelligence sources
- Known detection gaps and blind spots

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Review false positive rates and tune detection rules
- Quarterly: Update baselines for environmental changes
- Semi-annually: Validate detection effectiveness testing
- Annually: Complete reassessment of all detection capabilities
- Ad-hoc: When threat landscape changes or new attack vectors emerge

**Quality Assurance:**
Have SOC analysts and threat detection engineers validate assessments before
using results for compliance reporting or detection tuning decisions.

**CRITICAL - MTTD Target:**
Mean Time to Detect (MTTD) targets per ISMS-POL-A.8.16-S2.3:
- Critical events (Tier 1): ≤15 minutes
- High severity (Tier 2): ≤1 hour
- Medium severity (Tier 3): ≤24 hours

If detection capabilities cannot meet these SLAs, escalate immediately.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    import openpyxl
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.16.2"
WORKBOOK_NAME = "Baseline & Anomaly Detection Assessment"
CONTROL_ID = "A.8.16"
CONTROL_NAME = "Monitoring Activities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Output directory (WKBK/ sibling of SCR/)
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)



def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.16.2 specification
    sheets = [
        "Instructions & Legend",
        "1. Baseline Inventory",
        "2. Detection Rules",
        "3. MITRE ATT&CK Coverage",
        "4. Rule Performance",
        "5. Testing Validation",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "example_row": {
            "font": Font(name="Calibri", size=9, italic=True, color="808080"),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        },
        "severity_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "severity_high": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "severity_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "severity_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None,
            italic=style_dict["font"].italic if hasattr(style_dict["font"], 'italic') else False
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_planned': DataValidation(
            type="list",
            formula1='"Yes,No,Planned"',
            allow_blank=False
        ),
        'yes_no_partial_na': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,N/A"',
            allow_blank=False
        ),
        'baseline_type': DataValidation(
            type="list",
            formula1='"User Activity,System Activity,Network Traffic,Application Behaviour,Security Events,Resource Usage,Other"',
            allow_blank=False
        ),
        'baseline_status': DataValidation(
            type="list",
            formula1='"\u2705 Active,\u26A0\uFE0F Stale,\u274C Missing,↻ In Development"',
            allow_blank=False
        ),
        'update_frequency': DataValidation(
            type="list",
            formula1='"Daily,Weekly,Monthly,Quarterly,Annual,As Needed"',
            allow_blank=False
        ),
        'anomaly_detection': DataValidation(
            type="list",
            formula1='"Automated,Manual,Hybrid,None"',
            allow_blank=False
        ),
        'rule_type': DataValidation(
            type="list",
            formula1='"Correlation,Anomaly Detection,Threshold,Signature,Machine Learning,Threat Intel,Other"',
            allow_blank=False
        ),
        'severity': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Informational"',
            allow_blank=False
        ),
        'rule_status': DataValidation(
            type="list",
            formula1='"\u2705 Active,\u26A0\uFE0F Testing,\u274C Disabled,↻ Tuning"',
            allow_blank=False
        ),
        'tactic': DataValidation(
            type="list",
            formula1='"Reconnaissance,Resource Development,Initial Access,Execution,Persistence,Privilege Escalation,Defence Evasion,Credential Access,Discovery,Lateral Movement,Collection,Command and Control,Exfiltration,Impact"',
            allow_blank=False
        ),
        'coverage_status': DataValidation(
            type="list",
            formula1='"\u2705 Covered,\u26A0\uFE0F Partial,\u274C Not Covered,↻ Planned"',
            allow_blank=False
        ),
        'quality_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Fair,Poor"',
            allow_blank=False
        ),
        'tuning_status': DataValidation(
            type="list",
            formula1='"Optimized,Needs Tuning,Under Review,New Rule"',
            allow_blank=False
        ),
        'test_result': DataValidation(
            type="list",
            formula1='"Pass,Fail,Partial,Not Tested"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1='"\u2705 Compliant,\u26A0\uFE0F Partial,\u274C Non-Compliant,N/A"',
            allow_blank=False
        ),
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,None"',
            allow_blank=False
        ),
    }
    
    # Add all validations to worksheet
    for _dv in validations.values():
        ws.add_data_validation(_dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab in sequence (1–5).', '2. Document ALL baselines established for monitoring.', '3. Inventory ALL active detection/correlation rules.', '4. Map detection rules to MITRE ATT&CK framework.', '5. Assess rule performance (false positives, true positives).', '6. Document testing and validation activities.', '7. Review Summary Dashboard for coverage gaps.', '8. Gather evidence and list in Evidence Register.', '9. Obtain approvals via Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_baseline_inventory_sheet(ws, styles):
    """Create Baseline Inventory assessment sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:Q1")
    ws["A1"] = "1. BASELINE INVENTORY & MAINTENANCE"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:Q2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.2.1 - Establish and maintain behavioural baselines"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Are baselines established for critical systems, users, and network traffic to enable anomaly detection?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:Q3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Response dropdown
    ws["A4"] = "Overall Baseline Status:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    thin_s = Side(style="thin")
    ws["B4"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    ws.merge_cells("B4:E4")

    # Column Headers (Row 6)
    headers = [
        ("A", "Baseline Name/ID", 28),
        ("B", "Baseline Type", 22),
        ("C", "Scope/Target", 30),
        ("D", "Data Source", 25),
        ("E", "Baseline Period", 18),
        ("F", "Established Date", 16),
        ("G", "Last Updated", 16),
        ("H", "Update Frequency", 18),
        ("I", "Anomaly Detection Method", 25),
        ("J", "Threshold/Criteria", 25),
        ("K", "Alert Rules Linked", 20),
        ("L", "Baseline Status", 18),
        ("M", "Days Since Update", 16),
        ("N", "Compliance Status", 18),
        ("O", "Issues/Gaps", 30),
        ("P", "Owner", 20),
        ("Q", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "BL-USER-001",
        "User Activity",
        "Executive users",
        "Windows Event Logs",
        "30 days",
        "15.11.2024",
        "05.01.2025",
        "Monthly",
        "Automated",
        "3 std dev from mean",
        "RULE-105, RULE-203",
        "\u2705 Active",
        "2",
        "\u2705 Compliant",
        "None",
        "SOC Team",
        "High"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 1 sample + 50 empty, standard MAX-001)
    validations = create_base_validations(ws)

    for data_row in range(8, 58):
        for col_idx in range(1, 18):  # A to Q
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # Baseline Type
                validations['baseline_type'].add(cell)
            elif col_letter == 'H':  # Update Frequency
                validations['update_frequency'].add(cell)
            elif col_letter == 'I':  # Anomaly Detection
                validations['anomaly_detection'].add(cell)
            elif col_letter == 'L':  # Baseline Status
                validations['baseline_status'].add(cell)
            elif col_letter == 'M':  # Days Since Update (formula)
                cell.value = f'=IF(G{data_row}<>"",TODAY()-G{data_row},"")'
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            elif col_letter == 'N':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'Q':  # Priority
                validations['priority'].add(cell)

    # Baseline Statistics (Starting Row 59)
    row = 59
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "BASELINE STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"E{row}"] = "Count/Value"
    ws[f"J{row}"] = "Target"
    ws[f"N{row}"] = "Status"
    for col in ['A', 'E', 'J', 'N']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:D{row}")
    ws.merge_cells(f"E{row}:I{row}")
    ws.merge_cells(f"J{row}:M{row}")
    ws.merge_cells(f"N{row}:Q{row}")

    statistics = [
        ("Total Baselines Defined", "=COUNTA(A8:A85)", "N/A", ""),
        ("Active Baselines", "=COUNTIF(L8:L85,\"\u2705 Active\")", "All", ""),
        ("Stale Baselines (>90 days)", "=COUNTIF(M8:M85,\">90\")", "0", ""),
        ("Baselines by Type: User Activity", "=COUNTIF(B8:B85,\"User Activity\")", "Target", ""),
        ("Baselines by Type: System Activity", "=COUNTIF(B8:B85,\"System Activity\")", "Target", ""),
        ("Baselines by Type: Network Traffic", "=COUNTIF(B8:B85,\"Network Traffic\")", "Target", ""),
        ("Automated Anomaly Detection", "=COUNTIF(I8:I85,\"Automated\")", ">80%", ""),
        ("Baselines with Linked Rules", "=COUNTIFS(K8:K85,\"<>\")", "All", ""),
    ]

    row += 1
    for metric, formula, target, status in statistics:
        ws[f"A{row}"] = metric
        ws[f"E{row}"] = formula
        ws[f"J{row}"] = target
        ws[f"N{row}"] = status
        ws[f"N{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"N{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:D{row}")
        ws.merge_cells(f"E{row}:I{row}")
        ws.merge_cells(f"J{row}:M{row}")
        ws.merge_cells(f"N{row}:Q{row}")

        row += 1

    # Compliance Checklist (Starting Row 69)
    row = 69
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "BASELINE COMPLIANCE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Baselines established for Tier 1 (critical) systems",
        "Baselines established for Tier 2 (high) systems",
        "Baselines established for privileged user accounts",
        "Baselines established for network traffic patterns",
        "Baselines established for application behaviour",
        "Baseline establishment period adequate (min 30 days)",
        "Baseline update frequency defined and documented",
        "Baselines reviewed and updated regularly (max 90 days)",
        "Anomaly detection methods documented",
        "Anomaly thresholds defined and justified",
        "Alert rules linked to baseline deviations",
        "Baseline staleness monitored (alerts for >90 days)",
        "Baseline ownership assigned",
        "Baseline documentation maintained",
        "Baseline effectiveness reviewed quarterly",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:P{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:P{row}")
        
        # Status dropdown
        ws[f"Q{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style="thin")
        ws[f"Q{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"Q{row}"])


        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(Q71:Q85,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Table (after checklist score, add gap)
    row = score_row + 2
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "REFERENCE: BASELINE TYPES & RECOMMENDATIONS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_header_cols = [1, 5, 10, 14]  # A, E, J, N
    ref_headers = ["Baseline Type", "Example Metrics", "Recommended Period", "Update Frequency"]
    for col_idx, header in zip(ref_header_cols, ref_headers):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    ws.merge_cells(f"A{row}:D{row}")
    ws.merge_cells(f"E{row}:I{row}")
    ws.merge_cells(f"J{row}:M{row}")
    ws.merge_cells(f"N{row}:Q{row}")

    ref_data = [
        ("User Activity", "Logon times, locations, applications used", "30-60 days", "Monthly"),
        ("System Activity", "CPU, memory, disk I/O, process behaviour", "14-30 days", "Weekly"),
        ("Network Traffic", "Bandwidth, protocols, connection patterns", "30 days", "Monthly"),
        ("Application Behaviour", "Transaction rates, API calls, errors", "30-90 days", "Monthly"),
        ("Security Events", "Auth failures, privilege escalations", "90 days", "Quarterly"),
    ]

    row += 1
    for bl_type, metrics, period, frequency in ref_data:
        ws[f"A{row}"] = bl_type
        ws[f"E{row}"] = metrics
        ws[f"J{row}"] = period
        ws[f"N{row}"] = frequency
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:D{row}")
        ws.merge_cells(f"E{row}:I{row}")
        ws.merge_cells(f"J{row}:M{row}")
        ws.merge_cells(f"N{row}:Q{row}")

        row += 1


# ============================================================================
# SECTION 5: SHEET 3 - DETECTION RULES INVENTORY
# ============================================================================

def create_detection_rules_sheet(ws, styles):
    """Create Detection Rules Inventory sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:U1")
    ws["A1"] = "2. DETECTION RULES INVENTORY"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:U2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.2.2 - Detection rule coverage and effectiveness"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Are detection/correlation rules comprehensively covering security threats and regularly tuned for effectiveness?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:U3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Rule ID/Name", 30),
        ("B", "Rule Type", 22),
        ("C", "Description", 40),
        ("D", "Severity", 15),
        ("E", "Data Source(s)", 30),
        ("F", "Rule Logic Summary", 40),
        ("G", "MITRE Tactic", 25),
        ("H", "MITRE Technique", 25),
        ("I", "Created Date", 16),
        ("J", "Last Modified", 16),
        ("K", "Last Triggered", 16),
        ("L", "Trigger Count (30d)", 18),
        ("M", "True Positives (30d)", 18),
        ("N", "False Positives (30d)", 18),
        ("O", "Precision %", 15),
        ("P", "Rule Status", 18),
        ("Q", "Tuning Status", 18),
        ("R", "Owner", 20),
        ("S", "Compliance Status", 18),
        ("T", "Issues/Notes", 35),
        ("U", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "RULE-105",
        "Correlation",
        "Detect privilege escalation via Kerberos",
        "High",
        "Windows Event Logs",
        "4768 + 4769 + 4672 within 5 min",
        "Privilege Escalation",
        "T1078 - Valid Accounts",
        "10.08.2024",
        "15.12.2024",
        "05.01.2025",
        "12",
        "10",
        "2",
        "83%",
        "\u2705 Active",
        "Optimized",
        "SOC Team",
        "\u2705 Compliant",
        "None",
        "High"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-57: 50 rows for comprehensive rule inventory)
    validations = create_base_validations(ws)
    
    for data_row in range(8, 58):
        for col_idx in range(1, 22):  # A to U
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # Rule Type
                validations['rule_type'].add(cell)
            elif col_letter == 'D':  # Severity
                validations['severity'].add(cell)
            elif col_letter == 'G':  # MITRE Tactic
                validations['tactic'].add(cell)
            elif col_letter == 'O':  # Precision % (formula)
                cell.value = f'=IF(AND(M{data_row}<>"",N{data_row}<>""),ROUND(M{data_row}/(M{data_row}+N{data_row})*100,1)&"%","")'
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            elif col_letter == 'P':  # Rule Status
                validations['rule_status'].add(cell)
            elif col_letter == 'Q':  # Tuning Status
                validations['tuning_status'].add(cell)
            elif col_letter == 'S':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'U':  # Priority
                validations['priority'].add(cell)

    # Rule Statistics (Starting Row 60)
    row = 60
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "DETECTION RULE STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"F{row}"] = "Count/Value"
    ws[f"L{row}"] = "Target"
    ws[f"R{row}"] = "Status"
    for col in ['A', 'F', 'L', 'R']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:E{row}")
    ws.merge_cells(f"F{row}:K{row}")
    ws.merge_cells(f"L{row}:Q{row}")
    ws.merge_cells(f"R{row}:U{row}")

    statistics = [
        ("Total Active Rules", "=COUNTIF(P8:P57,\"\u2705 Active\")", "N/A", ""),
        ("Rules by Severity: Critical", "=COUNTIF(D8:D57,\"Critical\")", "Target", ""),
        ("Rules by Severity: High", "=COUNTIF(D8:D57,\"High\")", "Target", ""),
        ("Rules Triggered (30d)", "=SUM(L8:L57)", "N/A", ""),
        ("Average Precision %", "=AVERAGE(IF(O8:O57<>\"\",VALUE(LEFT(O8:O57,LEN(O8:O57)-1))))", ">70%", ""),
        ("Rules Needing Tuning", "=COUNTIF(Q8:Q57,\"Needs Tuning\")", "<10", ""),
        ("Rules Not Triggered (30d)", "=COUNTIF(L8:L57,0)", "<20%", ""),
        ("Rules Updated (30d)", "=COUNTIFS(J8:J57,\">\"&TODAY()-30)", "Target", ""),
    ]

    row += 1
    for metric, formula, target, status in statistics:
        ws[f"A{row}"] = metric
        ws[f"F{row}"] = formula
        ws[f"L{row}"] = target
        ws[f"R{row}"] = status
        ws[f"R{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"R{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:E{row}")
        ws.merge_cells(f"F{row}:K{row}")
        ws.merge_cells(f"L{row}:Q{row}")
        ws.merge_cells(f"R{row}:U{row}")

        row += 1

    # Compliance Checklist (Starting Row 72)
    row = 72
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "DETECTION RULE COMPLIANCE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Detection rules cover critical threats (malware, ransomware, data exfil)",
        "Rules aligned with MITRE ATT&CK framework",
        "Rules cover all MITRE tactics (Initial Access → Impact)",
        "Rule logic documented and peer-reviewed",
        "Rules tested before production deployment",
        "Rule performance tracked (TP, FP, precision)",
        "False positive rate acceptable (<30%)",
        "Rules tuned regularly (quarterly minimum)",
        "Inactive rules (not triggered 90+ days) reviewed",
        "Critical/High severity rules prioritised",
        "Rule ownership assigned",
        "Rules version controlled",
        "Rule changes require approval (change management)",
        "Rules tested against attack simulations",
        "Rule effectiveness reviewed quarterly",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:T{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:T{row}")
        
        # Status dropdown
        ws[f"U{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"U{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"U{row}"])
        

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(U74:U88,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

# ============================================================================
# SECTION 6: SHEET 4 - MITRE ATT&CK COVERAGE ASSESSMENT
# ============================================================================

def create_mitre_coverage_sheet(ws, styles):
    """Create MITRE ATT&CK Coverage mapping sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:N1")
    ws["A1"] = "3. MITRE ATT&CK FRAMEWORK COVERAGE"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:N2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.2.3 - Detection coverage mapped to threat framework"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Are detection capabilities mapped to MITRE ATT&CK framework with adequate coverage across all tactics?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:N3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "MITRE Tactic", 25),
        ("B", "MITRE Technique ID", 20),
        ("C", "Technique Name", 35),
        ("D", "Technique Description", 45),
        ("E", "Risk Level", 15),
        ("F", "Detection Rules", 30),
        ("G", "Rule Count", 12),
        ("H", "Data Sources Available", 30),
        ("I", "Coverage Status", 18),
        ("J", "Detection Quality", 18),
        ("K", "Last Tested", 16),
        ("L", "Test Result", 15),
        ("M", "Notes/Gaps", 35),
        ("N", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example rows showing MITRE tactics coverage
    example_data = [
        ("Initial Access", "T1078", "Valid Accounts", "Adversaries use valid accounts", "High", 
         "RULE-105, RULE-112", "2", "AD Logs, VPN Logs", "\u2705 Covered", "Good", 
         "20.12.2024", "Pass", "Working well", "High"),
        ("Execution", "T1059", "Command and Scripting Interpreter", "Execute commands via scripts", "High",
         "RULE-203, RULE-204", "2", "Process Logs, EDR", "\u2705 Covered", "Excellent",
         "18.12.2024", "Pass", "None", "High"),
        ("Persistence", "T1547", "Boot or Logon Autostart", "Persist via autostart", "Medium",
         "RULE-301", "1", "Registry, File System", "\u26A0\uFE0F Partial", "Fair",
         "10.12.2024", "Partial", "Needs more coverage", "Medium"),
    ]
    
    row = 7
    for data in example_data:
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["example_row"])
        row += 1

    # Data Entry Rows (Rows 10-59: 50 rows for technique coverage)
    validations = create_base_validations(ws)
    
    for data_row in range(10, 60):
        for col_idx in range(1, 15):  # A to N
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'A':  # MITRE Tactic
                validations['tactic'].add(cell)
            elif col_letter == 'E':  # Risk Level
                validations['severity'].add(cell)
            elif col_letter == 'G':  # Rule Count (formula)
                cell.value = f'=IF(F{data_row}<>"",LEN(F{data_row})-LEN(SUBSTITUTE(F{data_row},",",""))+1,0)'
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            elif col_letter == 'I':  # Coverage Status
                validations['coverage_status'].add(cell)
            elif col_letter == 'J':  # Detection Quality
                validations['quality_rating'].add(cell)
            elif col_letter == 'L':  # Test Result
                validations['test_result'].add(cell)
            elif col_letter == 'N':  # Priority
                validations['priority'].add(cell)

    # MITRE Tactics Summary (Starting Row 62)
    row = 62
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "MITRE ATT&CK TACTICS COVERAGE SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # MITRE 14 Tactics
    tactics_summary = [
        "Reconnaissance",
        "Resource Development",
        "Initial Access",
        "Execution",
        "Persistence",
        "Privilege Escalation",
        "Defence Evasion",
        "Credential Access",
        "Discovery",
        "Lateral Movement",
        "Collection",
        "Command and Control",
        "Exfiltration",
        "Impact",
    ]

    row += 1
    ws[f"A{row}"] = "Tactic"
    ws[f"C{row}"] = "Techniques Documented"
    ws[f"E{row}"] = "Covered"
    ws[f"G{row}"] = "Partial"
    ws[f"I{row}"] = "Not Covered"
    ws[f"L{row}"] = "Coverage %"
    for col in ['A', 'C', 'E', 'G', 'I', 'L']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:B{row}")
    ws.merge_cells(f"C{row}:D{row}")
    ws.merge_cells(f"E{row}:F{row}")
    ws.merge_cells(f"G{row}:H{row}")
    ws.merge_cells(f"I{row}:K{row}")
    ws.merge_cells(f"L{row}:N{row}")

    row += 1
    for tactic in tactics_summary:
        ws[f"A{row}"] = tactic
        ws[f"C{row}"] = f'=COUNTIF(A$10:A$59,"{tactic}")'
        ws[f"E{row}"] = f'=COUNTIFS(A$10:A$59,"{tactic}",I$10:I$59,"\u2705 Covered")'
        ws[f"G{row}"] = f'=COUNTIFS(A$10:A$59,"{tactic}",I$10:I$59,"\u26A0\uFE0F Partial")'
        ws[f"I{row}"] = f'=COUNTIFS(A$10:A$59,"{tactic}",I$10:I$59,"\u274C Not Covered")'
        ws[f"L{row}"] = f'=IF(C{row}>0,ROUND(E{row}/C{row}*100,1)&"%","N/A")'

        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:B{row}")
        ws.merge_cells(f"C{row}:D{row}")
        ws.merge_cells(f"E{row}:F{row}")
        ws.merge_cells(f"G{row}:H{row}")
        ws.merge_cells(f"I{row}:K{row}")
        ws.merge_cells(f"L{row}:N{row}")

        row += 1

    # Overall Coverage Stats
    row += 1
    ws[f"A{row}"] = "OVERALL COVERAGE"
    ws[f"C{row}"] = "=SUM(C64:C78)"
    ws[f"E{row}"] = "=SUM(E64:E78)"
    ws[f"G{row}"] = "=SUM(G64:G78)"
    ws[f"I{row}"] = "=SUM(I64:I78)"
    ws[f"L{row}"] = '=IF(C78>0,ROUND(E78/C78*100,1)&"%","N/A")'

    for col in ['A', 'C', 'E', 'G', 'I', 'L']:
        ws[f"{col}{row}"].font = Font(bold=True, size=11)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.merge_cells(f"A{row}:B{row}")
    ws.merge_cells(f"C{row}:D{row}")
    ws.merge_cells(f"E{row}:F{row}")
    ws.merge_cells(f"G{row}:H{row}")
    ws.merge_cells(f"I{row}:K{row}")
    ws.merge_cells(f"L{row}:N{row}")

    # Compliance Checklist (Starting Row 82)
    row = 82
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "MITRE ATT&CK COVERAGE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Detection rules mapped to MITRE ATT&CK framework",
        "All 14 MITRE tactics have at least partial coverage",
        "High-risk techniques (Initial Access, Execution) fully covered",
        "Coverage >60% across all tactics",
        "Coverage gaps identified and documented",
        "MITRE technique descriptions documented",
        "Required data sources identified for each technique",
        "Data source availability assessed",
        "Coverage tested via attack simulations",
        "Coverage reviewed quarterly",
        "New/emerging techniques evaluated",
        "Coverage gaps prioritised for remediation",
        "MITRE mapping maintained in detection rules",
        "Coverage metrics tracked over time",
        "MITRE Navigator export available",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:M{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:M{row}")
        
        # Status dropdown
        ws[f"N{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"N{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"N{row}"])
        

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(N84:N98,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 7: SHEET 5 - RULE PERFORMANCE & TUNING
# ============================================================================

def create_rule_performance_sheet(ws, styles):
    """Create Rule Performance & Tuning tracking sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:P1")
    ws["A1"] = "4. RULE PERFORMANCE & TUNING TRACKING"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:P2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.2.4 - Rule performance monitoring and optimization"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Are detection rules monitored for performance and regularly tuned to maintain effectiveness?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:P3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Performance Tracking Section (Starting Row 5)
    row = 5
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "RULE PERFORMANCE METRICS (Last 30 Days)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Rule ID/Name", 30),
        ("B", "Total Triggers", 15),
        ("C", "True Positives", 15),
        ("D", "False Positives", 15),
        ("E", "False Negatives", 15),
        ("F", "Precision %", 12),
        ("G", "Recall %", 12),
        ("H", "F1 Score", 12),
        ("I", "MTTD (minutes)", 15),
        ("J", "Tuning Actions", 30),
        ("K", "Last Tuned", 16),
        ("L", "Tuning Status", 18),
        ("M", "Performance Trend", 18),
        ("N", "Owner", 20),
        ("O", "Issues", 30),
        ("P", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "RULE-105",
        "12",
        "10",
        "2",
        "1",
        "83%",
        "91%",
        "0.87",
        "8",
        "Adjusted threshold",
        "15.12.2024",
        "Optimized",
        "Improving",
        "SOC Team",
        "None",
        "Medium"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-37: 30 rows)
    validations = create_base_validations(ws)
    
    for data_row in range(8, 38):
        for col_idx in range(1, 17):  # A to P
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply formulas and dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'F':  # Precision % formula: TP / (TP + FP)
                cell.value = f'=IF(AND(C{data_row}<>"",D{data_row}<>""),ROUND(C{data_row}/(C{data_row}+D{data_row})*100,1)&"%","")'
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            elif col_letter == 'G':  # Recall % formula: TP / (TP + FN)
                cell.value = f'=IF(AND(C{data_row}<>"",E{data_row}<>""),ROUND(C{data_row}/(C{data_row}+E{data_row})*100,1)&"%","")'
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            elif col_letter == 'H':  # F1 Score formula: 2 * (Precision * Recall) / (Precision + Recall)
                cell.value = f'=IF(AND(F{data_row}<>"",G{data_row}<>""),ROUND(2*(VALUE(LEFT(F{data_row},LEN(F{data_row})-1))/100)*(VALUE(LEFT(G{data_row},LEN(G{data_row})-1))/100)/((VALUE(LEFT(F{data_row},LEN(F{data_row})-1))/100)+(VALUE(LEFT(G{data_row},LEN(G{data_row})-1))/100)),2),"")'
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            elif col_letter == 'L':  # Tuning Status
                validations['tuning_status'].add(cell)
            elif col_letter == 'P':  # Priority
                validations['priority'].add(cell)

    # Performance Statistics (Starting Row 40)
    row = 40
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "PERFORMANCE STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"E{row}"] = "Value"
    ws[f"I{row}"] = "Target"
    ws[f"M{row}"] = "Status"
    for col in ['A', 'E', 'I', 'M']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:D{row}")
    ws.merge_cells(f"E{row}:H{row}")
    ws.merge_cells(f"I{row}:L{row}")
    ws.merge_cells(f"M{row}:P{row}")

    statistics = [
        ("Average Precision across all rules", "=AVERAGE(IF(F8:F37<>\"\",VALUE(LEFT(F8:F37,LEN(F8:F37)-1))))", ">70%", ""),
        ("Average Recall across all rules", "=AVERAGE(IF(G8:G37<>\"\",VALUE(LEFT(G8:G37,LEN(G8:G37)-1))))", ">80%", ""),
        ("Average F1 Score", "=AVERAGE(H8:H37)", ">0.75", ""),
        ("Rules with Precision <70%", "=COUNTIF(F8:F37,\"<70%\")", "<5", ""),
        ("Rules with high FP rate (>30%)", '=COUNTIF(F8:F37,"<70%")', "<5", ""),
        ("Rules requiring tuning", '=COUNTIF(L8:L37,"Needs Tuning")', "<10", ""),
        ("Rules tuned in last 30 days", "=COUNTIFS(K8:K37,\">\"&TODAY()-30)", "Target", ""),
        ("Average MTTD (minutes)", "=AVERAGE(I8:I37)", "<15", ""),
    ]

    row += 1
    for metric, formula, target, status in statistics:
        ws[f"A{row}"] = metric
        ws[f"E{row}"] = formula
        ws[f"I{row}"] = target
        ws[f"M{row}"] = status
        ws[f"M{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"M{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:D{row}")
        ws.merge_cells(f"E{row}:H{row}")
        ws.merge_cells(f"I{row}:L{row}")
        ws.merge_cells(f"M{row}:P{row}")

        row += 1

    # Tuning Actions Log (Starting Row 52)
    row = 52
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "RECENT TUNING ACTIONS LOG"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    tuning_headers = ["Date", "Rule ID", "Action Taken", "Reason", "Before Precision", "After Precision", "Performed By"]
    for col_idx, header in enumerate(tuning_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # 10 rows for tuning log
    for i in range(10):
        row += 1
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


    # Compliance Checklist (Starting Row 66)
    row = 66
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "RULE PERFORMANCE & TUNING CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Rule performance metrics collected (TP, FP, FN)",
        "Precision calculated for all active rules",
        "Recall calculated for all active rules",
        "F1 Score tracked to balance precision and recall",
        "MTTD tracked for critical/high severity rules",
        "Rules with precision <70% identified for tuning",
        "False positive rate acceptable (<30%)",
        "False negative testing performed",
        "Tuning actions documented and tracked",
        "Rules tuned regularly (quarterly minimum)",
        "Performance trends monitored over time",
        "High-priority rules tuned within 30 days of issues",
        "Tuning effectiveness measured (before/after metrics)",
        "Rule owners assigned and accountable",
        "Performance metrics reviewed in quarterly meetings",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:O{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:O{row}")
        
        # Status dropdown
        ws[f"P{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"P{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"P{row}"])
        

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(P68:P82,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 8: SHEET 6 - TESTING & VALIDATION
# ============================================================================

def create_testing_validation_sheet(ws, styles):
    """Create Testing & Validation tracking sheet."""
    
    ws.row_dimensions[1].height = 35
    # Header
    ws.merge_cells("A1:N1")
    ws["A1"] = "5. DETECTION TESTING & VALIDATION"
    apply_style(ws["A1"], styles["header"])

    ws.merge_cells("A2:N2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.2.5 - Detection rule testing and validation"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Are detection rules tested and validated to ensure they detect intended threats?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:N3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Test ID", 15),
        ("B", "Rule ID/Name", 30),
        ("C", "Test Date", 16),
        ("D", "Test Type", 20),
        ("E", "Test Scenario", 40),
        ("F", "Expected Result", 35),
        ("G", "Actual Result", 35),
        ("H", "Test Result", 15),
        ("I", "Detection Time", 15),
        ("J", "Issues Identified", 35),
        ("K", "Remediation Action", 35),
        ("L", "Retested", 15),
        ("M", "Tested By", 20),
        ("N", "Notes", 35),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "TEST-001",
        "RULE-105",
        "20.12.2024",
        "Attack Simulation",
        "Kerberos Golden Ticket attack",
        "Alert triggered within 5 min",
        "Alert triggered in 3 min",
        "Pass",
        "3 min",
        "None",
        "N/A",
        "Yes",
        "SOC Team",
        "Working as expected"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-37: 30 rows)
    validations = create_base_validations(ws)
    
    test_type_validation = DataValidation(
        type="list",
        formula1='"Attack Simulation,Purple Team,Unit Test,Integration Test,Regression Test,Penetration Test"',
        allow_blank=False
    )
    ws.add_data_validation(test_type_validation)

    for data_row in range(8, 38):
        # Auto-generate Test ID
        ws[f"A{data_row}"] = f'="TEST-"&TEXT(ROW()-7,"000")'
        ws[f"A{data_row}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_a = Side(style="thin")
        ws[f"A{data_row}"].border = Border(left=thin_a, right=thin_a, top=thin_a, bottom=thin_a)

        for col_idx in range(2, 15):  # B to N
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            # Apply dropdowns
            col_letter = get_column_letter(col_idx)
            if col_letter == 'D':  # Test Type
                test_type_validation.add(cell)
            elif col_letter == 'H':  # Test Result
                validations['test_result'].add(cell)
            elif col_letter == 'L':  # Retested
                validations['yes_no'].add(cell)
        


    # Test Statistics (Starting Row 40)
    row = 40
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "TESTING STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"E{row}"] = "Count/Value"
    ws[f"I{row}"] = "Target"
    ws[f"L{row}"] = "Status"
    for col in ['A', 'E', 'I', 'L']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:D{row}")
    ws.merge_cells(f"E{row}:H{row}")
    ws.merge_cells(f"I{row}:K{row}")
    ws.merge_cells(f"L{row}:N{row}")

    statistics = [
        ("Total Tests Performed", "=COUNTA(B8:B37)", "N/A", ""),
        ("Tests Passed", '=COUNTIF(H8:H37,"Pass")', "Target", ""),
        ("Tests Failed", '=COUNTIF(H8:H37,"Fail")', "<5%", ""),
        ("Tests Partial", '=COUNTIF(H8:H37,"Partial")', "Review", ""),
        ("Pass Rate %", '=IF(COUNTA(H8:H37)>0,ROUND(COUNTIF(H8:H37,"Pass")/COUNTA(H8:H37)*100,1)&"%","N/A")', ">90%", ""),
        ("Average Detection Time", "=AVERAGE(IF(I8:I37<>\"\",VALUE(LEFT(I8:I37,FIND(\" \",I8:I37)-1))))", "<10 min", ""),
        ("Tests Requiring Retest", '=COUNTIF(L8:L37,"No")', "0", ""),
        ("Rules Not Tested (90+ days)", "[Manual Assessment]", "0", ""),
    ]

    row += 1
    for metric, formula, target, status in statistics:
        ws[f"A{row}"] = metric
        ws[f"E{row}"] = formula
        ws[f"I{row}"] = target
        ws[f"L{row}"] = status
        ws[f"L{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"L{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:D{row}")
        ws.merge_cells(f"E{row}:H{row}")
        ws.merge_cells(f"I{row}:K{row}")
        ws.merge_cells(f"L{row}:N{row}")

        row += 1

    # Compliance Checklist (Starting Row 52)
    row = 52
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "TESTING & VALIDATION CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "All critical/high severity rules tested before production",
        "Testing performed using realistic attack scenarios",
        "Attack simulations (Purple Team exercises) conducted",
        "Detection time measured and within acceptable range",
        "Unit tests performed for individual rule logic",
        "Integration tests performed for multi-event correlation",
        "Regression testing after rule modifications",
        "Test results documented with evidence",
        "Failed tests result in rule improvement",
        "Rules tested quarterly minimum",
        "MITRE techniques tested via attack simulations",
        "False negative testing performed",
        "Detection gaps identified and documented",
        "Testing schedule maintained and followed",
        "Test results reviewed by SOC management",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:M{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:M{row}")
        
        # Status dropdown
        ws[f"N{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin_s = Side(style='thin')
        ws[f"N{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        validations['compliance_status'].add(ws[f"N{row}"])
        

        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(N54:N68,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with TABLE 1/2/3 Gold Standard layout."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: A1:G1 merged 003366 header
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{WORKBOOK_NAME.upper()} \u2014 SUMMARY DASHBOARD'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle
    ws['A2'] = f'Summary Dashboard — {CONTROL_NAME} | {GENERATED_DATE}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='left')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.freeze_panes = 'A4'

    # TABLE 1: COMPLIANCE OVERVIEW
    ws.merge_cells('A4:G4')
    ws['A4'] = 'TABLE 1: COMPLIANCE OVERVIEW'
    ws['A4'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A4'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A4'].border = border

    # Row 5: headers
    t1_headers = ['Assessment Area', 'Total Items', 'Compliant', 'Partial', 'Non-Compliant', 'N/A', 'Compliance %']
    for col_idx, header in enumerate(t1_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='000000')
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Rows 6-10: data rows
    ws.cell(row=6, column=1, value='1. Baseline Inventory').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=6, column=2, value='=SUM(C6:F6)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=3, value='=COUNTIF(\'1. Baseline Inventory\'!Q71:Q85,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=4, value='=COUNTIF(\'1. Baseline Inventory\'!Q71:Q85,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=5, value='=COUNTIF(\'1. Baseline Inventory\'!Q71:Q85,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=6, value='=COUNTIF(\'1. Baseline Inventory\'!Q71:Q85,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7, value='=IF((B6-F6)=0,0,C6/(B6-F6))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=6, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=6, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=6, column=col_idx).border = border

    ws.cell(row=7, column=1, value='2. Detection Rules').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=7, column=2, value='=SUM(C7:F7)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=3, value='=COUNTIF(\'2. Detection Rules\'!U74:U88,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=4, value='=COUNTIF(\'2. Detection Rules\'!U74:U88,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=5, value='=COUNTIF(\'2. Detection Rules\'!U74:U88,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=6, value='=COUNTIF(\'2. Detection Rules\'!U74:U88,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7, value='=IF((B7-F7)=0,0,C7/(B7-F7))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=7, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=7, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=7, column=col_idx).border = border

    ws.cell(row=8, column=1, value='3. MITRE ATT&CK Coverage').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=8, column=2, value='=SUM(C8:F8)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=3, value='=COUNTIF(\'3. MITRE ATT&CK Coverage\'!N84:N98,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=4, value='=COUNTIF(\'3. MITRE ATT&CK Coverage\'!N84:N98,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=5, value='=COUNTIF(\'3. MITRE ATT&CK Coverage\'!N84:N98,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=6, value='=COUNTIF(\'3. MITRE ATT&CK Coverage\'!N84:N98,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7, value='=IF((B8-F8)=0,0,C8/(B8-F8))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=8, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=8, column=col_idx).border = border

    ws.cell(row=9, column=1, value='4. Rule Performance').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=9, column=2, value='=SUM(C9:F9)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=3, value='=COUNTIF(\'4. Rule Performance\'!P68:P82,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=4, value='=COUNTIF(\'4. Rule Performance\'!P68:P82,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=5, value='=COUNTIF(\'4. Rule Performance\'!P68:P82,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=6, value='=COUNTIF(\'4. Rule Performance\'!P68:P82,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7, value='=IF((B9-F9)=0,0,C9/(B9-F9))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=9, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=9, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=9, column=col_idx).border = border

    ws.cell(row=10, column=1, value='5. Testing Validation').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=10, column=2, value='=SUM(C10:F10)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=3, value='=COUNTIF(\'5. Testing Validation\'!N54:N68,"✅ Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=4, value='=COUNTIF(\'5. Testing Validation\'!N54:N68,"⚠️ Partial")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=5, value='=COUNTIF(\'5. Testing Validation\'!N54:N68,"❌ Non-Compliant")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=6, value='=COUNTIF(\'5. Testing Validation\'!N54:N68,"N/A")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7, value='=IF((B10-F10)=0,0,C10/(B10-F10))').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=10, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=10, column=7).number_format = '0.0%'
    for col_idx in range(1, 8):
        ws.cell(row=10, column=col_idx).border = border

    # Row 11: TOTAL
    ws.cell(row=11, column=1, value='TOTAL').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=1).border = border
    for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F'], start=2):
        ws.cell(row=11, column=col_idx, value=f'=SUM({col_letter}6:{col_letter}10)').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=11, column=col_idx).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws.cell(row=11, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=11, column=col_idx).border = border
    ws.cell(row=11, column=7, value='=IF((B11-F11)=0,0,C11/(B11-F11))').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=11, column=7).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.cell(row=11, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=11, column=7).number_format = '0.0%'
    ws.cell(row=11, column=7).border = border

    # TABLE 2: KEY METRICS
    ws.merge_cells('A13:G13')
    ws['A13'] = 'TABLE 2: KEY METRICS'
    ws['A13'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A13'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A13'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A13'].border = border
    ws.merge_cells('A14:E14')
    ws['A14'] = 'Metric'
    ws['A14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A14'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A14'].border = border
    ws.merge_cells('F14:G14')
    ws['F14'] = 'Value'
    ws['F14'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F14'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F14'].border = border
    yllw_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws.merge_cells('A15:E15')
    ws.cell(row=15, column=1, value='Active Baselines').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=15, column=1).border = border
    ws.merge_cells('F15:G15')
    ws.cell(row=15, column=6, value='=COUNTIF(\'1. Baseline Inventory\'!L8:L200,"✅ Active")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=15, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=15, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=15, column=col_idx).border = border

    ws.merge_cells('A16:E16')
    ws.cell(row=16, column=1, value='Stale Baselines (>90 days)').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=16, column=1).border = border
    ws.merge_cells('F16:G16')
    ws.cell(row=16, column=6, value='=COUNTIF(\'1. Baseline Inventory\'!L8:L200,"⚠️ Stale")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=16, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=16, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=16, column=col_idx).border = border

    ws.merge_cells('A17:E17')
    ws.cell(row=17, column=1, value='Baselines \u2014 Missing/Unconfigured').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=17, column=1).border = border
    ws.merge_cells('F17:G17')
    ws.cell(row=17, column=6, value='=COUNTIF(\'1. Baseline Inventory\'!L8:L200,"\u274C Missing")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=17, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=17, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=17, column=col_idx).border = border

    ws.merge_cells('A18:E18')
    ws.cell(row=18, column=1, value='Active Detection Rules').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=18, column=1).border = border
    ws.merge_cells('F18:G18')
    ws.cell(row=18, column=6, value='=COUNTIF(\'2. Detection Rules\'!P8:P200,"✅ Active")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=18, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=18, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=18, column=col_idx).border = border

    ws.merge_cells('A19:E19')
    ws.cell(row=19, column=1, value='Rules Needing Tuning').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=19, column=1).border = border
    ws.merge_cells('F19:G19')
    ws.cell(row=19, column=6, value='=COUNTIF(\'2. Detection Rules\'!Q8:Q200,"Needs Tuning")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=19, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=19, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=19, column=col_idx).border = border

    ws.merge_cells('A20:E20')
    ws.cell(row=20, column=1, value='Critical/High Severity Rules').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=20, column=1).border = border
    ws.merge_cells('F20:G20')
    ws.cell(row=20, column=6, value='=COUNTIF(\'2. Detection Rules\'!D8:D200,"Critical")+COUNTIF(\'2. Detection Rules\'!D8:D200,"High")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=20, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=20, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=20, column=col_idx).border = border

    ws.merge_cells('A21:E21')
    ws.cell(row=21, column=1, value='MITRE Techniques — Covered').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=21, column=1).border = border
    ws.merge_cells('F21:G21')
    ws.cell(row=21, column=6, value='=COUNTIF(\'3. MITRE ATT&CK Coverage\'!I8:I200,"✅ Covered")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=21, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=21, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=21, column=col_idx).border = border

    ws.merge_cells('A22:E22')
    ws.cell(row=22, column=1, value='Critical MITRE Techniques — Not Covered').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=22, column=1).border = border
    ws.merge_cells('F22:G22')
    ws.cell(row=22, column=6, value='=COUNTIFS(\'3. MITRE ATT&CK Coverage\'!E8:E200,"Critical",\'3. MITRE ATT&CK Coverage\'!I8:I200,"❌ Not Covered")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=22, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=22, column=6).border = border
    for col_idx in range(1, 8):
        ws.cell(row=22, column=col_idx).border = border

    # TABLE 3: CRITICAL FINDINGS
    ws.merge_cells('A24:G24')
    ws['A24'] = 'TABLE 3: CRITICAL FINDINGS'
    ws['A24'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A24'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A24'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A24'].border = border
    ws.merge_cells('A25:E25')
    ws['A25'] = 'Category / Finding'
    ws['A25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['A25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['A25'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A25'].border = border
    ws['F25'] = 'Count'
    ws['F25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['F25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F25'].border = border
    ws['G25'] = 'Action'
    ws['G25'].font = Font(name='Calibri', size=10, bold=True, color='000000')
    ws['G25'].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws['G25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G25'].border = border

    ws.merge_cells('A26:E26')
    ws.cell(row=26, column=1, value='Detection Rules \u2014 Disabled').fill = yllw_fill
    ws.cell(row=26, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=26, column=1).border = border
    ws.cell(row=26, column=6, value='=COUNTIF(\'2. Detection Rules\'!P8:P200,"\u274C Disabled")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=26, column=6).fill = yllw_fill
    ws.cell(row=26, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=26, column=6).border = border
    ws.cell(row=26, column=7, value='').fill = yllw_fill
    ws.cell(row=26, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=26, column=col_idx).fill = yllw_fill
        ws.cell(row=26, column=col_idx).border = border

    ws.merge_cells('A27:E27')
    ws.cell(row=27, column=1, value='Rules Needing Tuning').fill = yllw_fill
    ws.cell(row=27, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=27, column=1).border = border
    ws.cell(row=27, column=6, value='=COUNTIF(\'2. Detection Rules\'!Q8:Q200,"Needs Tuning")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=27, column=6).fill = yllw_fill
    ws.cell(row=27, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=27, column=6).border = border
    ws.cell(row=27, column=7, value='').fill = yllw_fill
    ws.cell(row=27, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=27, column=col_idx).fill = yllw_fill
        ws.cell(row=27, column=col_idx).border = border

    ws.merge_cells('A28:E28')
    ws.cell(row=28, column=1, value='Stale Baselines').fill = yllw_fill
    ws.cell(row=28, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=28, column=1).border = border
    ws.cell(row=28, column=6, value='=COUNTIF(\'1. Baseline Inventory\'!L8:L200,"⚠️ Stale")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=28, column=6).fill = yllw_fill
    ws.cell(row=28, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=28, column=6).border = border
    ws.cell(row=28, column=7, value='').fill = yllw_fill
    ws.cell(row=28, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=28, column=col_idx).fill = yllw_fill
        ws.cell(row=28, column=col_idx).border = border

    ws.merge_cells('A29:E29')
    ws.cell(row=29, column=1, value='MITRE Techniques — Not Covered').fill = yllw_fill
    ws.cell(row=29, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=29, column=1).border = border
    ws.cell(row=29, column=6, value='=COUNTIF(\'3. MITRE ATT&CK Coverage\'!I8:I200,"❌ Not Covered")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=29, column=6).fill = yllw_fill
    ws.cell(row=29, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=29, column=6).border = border
    ws.cell(row=29, column=7, value='').fill = yllw_fill
    ws.cell(row=29, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=29, column=col_idx).fill = yllw_fill
        ws.cell(row=29, column=col_idx).border = border

    ws.merge_cells('A30:E30')
    ws.cell(row=30, column=1, value='Critical MITRE Gaps').fill = yllw_fill
    ws.cell(row=30, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=30, column=1).border = border
    ws.cell(row=30, column=6, value='=COUNTIFS(\'3. MITRE ATT&CK Coverage\'!E8:E200,"Critical",\'3. MITRE ATT&CK Coverage\'!I8:I200,"❌ Not Covered")').font = Font(name='Calibri', size=10, color='000000')
    ws.cell(row=30, column=6).fill = yllw_fill
    ws.cell(row=30, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=30, column=6).border = border
    ws.cell(row=30, column=7, value='').fill = yllw_fill
    ws.cell(row=30, column=7).border = border
    for col_idx in range(1, 8):
        ws.cell(row=30, column=col_idx).fill = yllw_fill
        ws.cell(row=30, column=col_idx).border = border

    # Row 31: instruction
    ws.merge_cells('A31:G31')
    ws['A31'] = 'Filter source sheets using AutoFilter — see column headers above'
    ws['A31'].font = Font(name='Calibri', size=9, italic=True, color='003366')
    ws['A31'].alignment = Alignment(horizontal='left', vertical='center')

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet — Gold Standard 8-column format."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: merged title, 003366 fill, white bold 14pt + borders all 8 cells
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _c in range(1, 9):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: generic subtitle — italic, left-aligned
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty separator

    # Row 4: column headers — 003366 fill, white bold
    er_headers = [
        ('Evidence ID', 15),
        ('Assessment Area', 25),
        ('Evidence Type', 22),
        ('Description', 40),
        ('Location/Path', 45),
        ('Date Collected', 16),
        ('Collected By', 20),
        ('Verification Status', 22),
    ]
    for col_idx, (header, width) in enumerate(er_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Log Export,Documentation,Report,Network scan,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    ws.add_data_validation(status_dv)

    type_dv.add("C6:C105")
    status_dv.add("H6:H105")

    # Row 5: F2F2F2 grey sample row — fully populated example
    sample = [
        'EV-001',
        'Baseline Detection',
        'Log Export',
        'Network baseline detection threshold rules export from QRadar',
        '\\\\fileserver\\isms\\evidence\\monitoring\\baseline-rules-qradar.csv',
        '2024-03-01',
        'Sarah Mitchell',
        'Verified',
    ]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.font = Font(name="Calibri", size=10)

    # Rows 6-105: 100 empty FFFFCC data rows (no pre-populated IDs)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Freeze pane at A5 (freeze rows 1–4)
    ws.freeze_panes = "A5"

def _as2_border():
    """Return a fresh thin border object."""
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def create_approval_sheet(ws, styles):
    """Create Approval Sign-Off sheet matching golden standard."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _bm(ref):
        ws[ref].border = _as2_border()

    # ── Header (Row 1) — merged A1:E1 ──
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = _as2_border()
    ws.row_dimensions[1].height = 35

    # ── Subtitle (Row 2) ──
    ws.merge_cells("A2:E2")
    ws["A2"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = _as2_border()
    # ── Freeze pane at A3 ──
    ws.freeze_panes = "A3"

    # ── ASSESSMENT SUMMARY banner (Row 3) ──
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLETED BY"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as2_border()
    # ── Summary fields (Row 4+) ──
    summary_fields = [
        ("Name:", ""),
        ("Title:", ""),
        ("Date:", ""),
        ("Signature:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as2_border()
        _bm(f"B{row}")
        row += 1

    # ── 2 more Approver sections ──
    approvers = [
        ("REVIEWED BY", "4472C4"),
        ("APPROVED BY", "003366"),
    ]

    row += 1  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = _as2_border()
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = _as2_border()
            _bm(f"B{row}")
            row += 1
        row += 1  # gap between sections

    # ── FINAL DECISION ──
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "FINAL DECISION"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as2_border()
    row += 1

    ws[f"A{row}"] = "Decision:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = _as2_border()
    _bm(f"B{row}")

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)

    decision_dv.add(ws[f"B{row}"])

    # ── NEXT REVIEW ──
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = _as2_border()
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = _as2_border()
        _bm(f"B{row}")
        row += 1

    # ── Column widths ──
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20


# ============================================================================
# SECTION 11: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "It doesn't matter how beautiful your theory is, it doesn't 
    matter how smart you are. If it doesn't agree with experiment, it's wrong."
    - Richard Feynman
    
    Detection rules that look good on paper but fail in practice are worse 
    than no detection at all. They give false confidence while attackers 
    walk right through. Test. Measure. Improve. That's the way.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.16.2 - Baseline & Detection Rules Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities")
    logger.info("=" * 78)
    logger.info("\n>>> Systems Engineering: Test-Driven Detection")
    logger.info(">>> Comprehensive: Baselines, Rules, MITRE, Performance, Testing")
    logger.info(">>> Audit-Ready: 75 compliance checkpoints, 100 evidence entries")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("\u2705 Workbook created with 9 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")
    
    logger.info("  [1/9] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  \u2705 Instructions complete")

    logger.info("  [2/9] Creating Baseline Inventory...")
    create_baseline_inventory_sheet(wb["1. Baseline Inventory"], styles)
    logger.info("  \u2705 Baseline tracking complete (30 baseline rows, 15 checks)")

    logger.info("  [3/9] Creating Detection Rules Inventory...")
    create_detection_rules_sheet(wb["2. Detection Rules"], styles)
    logger.info("  \u2705 Rule inventory complete (50 rule rows, 15 checks)")

    logger.info("  [4/9] Creating MITRE ATT&CK Coverage...")
    create_mitre_coverage_sheet(wb["3. MITRE ATT&CK Coverage"], styles)
    logger.info("  \u2705 MITRE mapping complete (50 technique rows, 14 tactics)")

    logger.info("  [5/9] Creating Rule Performance & Tuning...")
    create_rule_performance_sheet(wb["4. Rule Performance"], styles)
    logger.info("  \u2705 Performance tracking complete (30 rule rows, metrics)")

    logger.info("  [6/9] Creating Testing & Validation...")
    create_testing_validation_sheet(wb["5. Testing Validation"], styles)
    logger.info("  \u2705 Testing log complete (30 test rows, 15 checks)")

    logger.info("  [8/9] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  \u2705 Evidence register complete (100 evidence rows)")

    logger.info("  [7/9] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  \u2705 Dashboard complete (consolidated compliance view)")

    logger.info("  [9/9] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  \u2705 Approval workflow complete (3-level sign-off)")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.16.2_Baseline_Detection_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = _wkbk_dir / OUTPUT_FILENAME
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"\u2705 SUCCESS: {output_path}")
    except Exception as e:
        logger.error(f"\u274C ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info("WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets:")
    logger.info("  \u2022 1. Baseline Inventory (30 baseline rows, staleness tracking)")
    logger.info("  \u2022 2. Detection Rules (50 rule rows, performance metrics)")
    logger.info("  \u2022 3. MITRE ATT&CK Coverage (50 techniques, 14 tactics)")
    logger.info("  \u2022 4. Rule Performance & Tuning (precision, recall, F1 score)")
    logger.error("  \u2022 5. Testing & Validation (30 test rows, pass/fail tracking)")
    logger.info("\n>>> Consolidation & Governance:")
    logger.info("  \u2022 Summary Dashboard (overall compliance, key metrics)")
    logger.info("  \u2022 Evidence Register (100 evidence entries)")
    logger.info("  \u2022 Approval Sign-Off (4-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info(">>> ASSESSMENT CAPABILITIES:")
    logger.info("  \u2022 75 compliance checkpoint items across 5 assessment areas")
    logger.info("  \u2022 30 baseline tracking rows with staleness monitoring")
    logger.info("  \u2022 50 detection rule inventory rows")
    logger.info("  \u2022 50 MITRE ATT&CK technique mapping rows")
    logger.info("  \u2022 14 MITRE tactics coverage summary")
    logger.info("  \u2022 30 rule performance tracking rows (TP, FP, precision, recall)")
    logger.info("  \u2022 30 test case tracking rows")
    logger.info("  \u2022 Automated compliance % calculations")
    logger.info("  \u2022 Automated precision/recall/F1 formulas")
    logger.info("\n" + "─" * 78)
    logger.info(">>> KEY FEATURES:")
    logger.info("  \u2705 Baseline establishment and staleness tracking")
    logger.info("  \u2705 Comprehensive rule inventory with performance metrics")
    logger.info("  \u2705 MITRE ATT&CK framework mapping (14 tactics)")
    logger.info("  \u2705 Rule performance tracking (precision, recall, F1 score)")
    logger.info("  \u2705 Detection testing and validation log")
    logger.info("  \u2705 Automated calculations for TP, FP, precision, recall")
    logger.info("  \u2705 MITRE coverage % by tactic")
    logger.info("  \u2705 Tuning action tracking")
    logger.info("  \u2705 Multi-level approval workflow")
    logger.info("  \u2705 Quarterly review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(">>> NEXT STEPS:")
    logger.info("  1. Document all established baselines")
    logger.info("  2. Inventory ALL active detection/correlation rules")
    logger.info("  3. Map rules to MITRE ATT&CK framework")
    logger.info("  4. Track rule performance (TP, FP, precision, recall)")
    logger.info("  5. Document testing and validation activities")
    logger.info("  6. Review Summary Dashboard for gaps")
    logger.info("  7. Address rules with precision <70%")
    logger.info("  8. Improve MITRE coverage to >60%")
    logger.info("  9. Document evidence in Evidence Register")
    logger.info("  10. Obtain final approval via Approval Sign-Off")
    logger.info("\n>>> PRO TIP:")
    logger.info("  Detection without testing is just hopeful thinking.")
    logger.info("  Your rules might look beautiful, but if they don't catch")
    logger.info("  attackers in practice, they're worse than useless.")
    logger.info("  Test against MITRE techniques. Measure precision and recall.")
    logger.info("  Tune based on evidence, not gut feeling.")
    logger.info("  That's Systems Engineering applied to detection.")
    logger.info("\n" + "=" * 78)
    logger.info('\n"It doesn\'t matter how beautiful your theory is...')
    logger.info("\n>>> This is not cargo cult ISMS. This is test-driven detection.")
    logger.info(">>> We MEASURE effectiveness. We VALIDATE with simulations.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
