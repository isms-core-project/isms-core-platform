#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.30.3 - Security Testing and Acceptance
================================================================================

ISO/IEC 27001:2022 Control A.8.30: Outsourced Development
Assessment Domain 3 of 4: Security Testing and Acceptance

This script generates a comprehensive Excel assessment workbook for tracking
security testing activities and acceptance criteria for outsourced development
deliverables.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.30.3"
WORKBOOK_NAME = "Security Testing and Acceptance"
CONTROL_ID = "A.8.30"
CONTROL_NAME = "Outsourced Development"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.8.30.3 - Security Testing and Acceptance"],
        [""],
        ["PURPOSE"],
        ["This workbook tracks security testing activities and acceptance criteria"],
        ["for outsourced development deliverables per ISMS-POL-A.8.30."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Deliverable Inventory - All deliverables from outsourced development"],
        ["3. Code Review Tracking - Internal code review records"],
        ["4. Security Testing - SAST, DAST, SCA, penetration test results"],
        ["5. SBOM Management - Software Bill of Materials tracking"],
        ["6. Acceptance Sign-off - Final acceptance criteria verification"],
        [""],
        ["WORKFLOW"],
        ["1. Register deliverable in inventory"],
        ["2. Initiate code review"],
        ["3. Execute SAST scan"],
        ["4. Execute SCA scan (generate SBOM)"],
        ["5. Execute DAST scan (for web apps)"],
        ["6. Penetration test (for Critical projects)"],
        ["7. Triage and remediate findings"],
        ["8. Retest Critical/High findings"],
        ["9. Verify acceptance criteria"],
        ["10. Obtain sign-off"],
        [""],
        ["DATA VALIDATION"],
        ["- Yellow cells are input fields"],
        ["- Dropdown lists enforce valid values"],
        ["- Deliverable_ID must be unique (DEL-XXXX format)"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "WORKFLOW", "DATA VALIDATION"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 80


def create_deliverable_inventory_sheet(ws):
    """Create the Deliverable Inventory sheet."""
    ws.title = "Deliverable Inventory"

    headers = [
        "Deliverable_ID", "Contract_ID", "Vendor_ID", "Deliverable_Name",
        "Deliverable_Type", "Project_Classification", "Planned_Delivery",
        "Actual_Delivery", "Code_Review_Status", "Security_Test_Status",
        "SBOM_Received", "Acceptance_Status", "Acceptance_Date", "Accepted_By"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Application,Module,Component,Library,API"')
    ws.add_data_validation(type_dv)
    type_dv.add('E2:E100')

    class_dv = DataValidation(type="list", formula1='"Critical,High,Standard"')
    ws.add_data_validation(class_dv)
    class_dv.add('F2:F100')

    review_dv = DataValidation(type="list", formula1='"Pending,In Progress,Complete,N/A"')
    ws.add_data_validation(review_dv)
    review_dv.add('I2:I100')

    test_dv = DataValidation(type="list", formula1='"Pending,In Progress,Complete"')
    ws.add_data_validation(test_dv)
    test_dv.add('J2:J100')

    sbom_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(sbom_dv)
    sbom_dv.add('K2:K100')

    accept_dv = DataValidation(type="list", formula1='"Pending,Accepted,Rejected,Conditional"')
    ws.add_data_validation(accept_dv)
    accept_dv.add('L2:L100')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Set column widths
    widths = [15, 12, 12, 35, 15, 18, 15, 15, 18, 18, 15, 15, 15, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_code_review_sheet(ws):
    """Create the Code Review Tracking sheet."""
    ws.title = "Code Review Tracking"

    headers = [
        "Review_ID", "Deliverable_ID", "Review_Type", "Review_Date",
        "Reviewer", "Reviewer_Role", "Files_Reviewed", "Security_Findings",
        "Critical_Findings", "High_Findings", "Medium_Findings", "Low_Findings",
        "Review_Result", "Findings_Reference", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Peer Review,Security Review,Architecture Review"')
    ws.add_data_validation(type_dv)
    type_dv.add('C2:C100')

    role_dv = DataValidation(type="list", formula1='"Developer,Security Team,Security Architect"')
    ws.add_data_validation(role_dv)
    role_dv.add('F2:F100')

    result_dv = DataValidation(type="list", formula1='"Approved,Approved with Findings,Rejected"')
    ws.add_data_validation(result_dv)
    result_dv.add('M2:M100')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 18, 12, 20, 18, 15, 15, 15, 12, 15, 12, 20, 35, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_security_testing_sheet(ws):
    """Create the Security Testing Results sheet."""
    ws.title = "Security Testing"

    headers = [
        "Test_ID", "Deliverable_ID", "Test_Type", "Test_Tool", "Test_Date",
        "Tester", "Scope", "Total_Findings", "Critical_Findings", "High_Findings",
        "Medium_Findings", "Low_Findings", "False_Positives", "Findings_Remediated",
        "Findings_Outstanding", "Report_Reference", "Retest_Required",
        "Retest_Date", "Retest_Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"SAST,DAST,SCA,Penetration Test,Manual Review"')
    ws.add_data_validation(type_dv)
    type_dv.add('C2:C100')

    retest_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(retest_dv)
    retest_dv.add('Q2:Q100')

    retest_status_dv = DataValidation(type="list", formula1='"Pending,Passed,Failed,N/A"')
    ws.add_data_validation(retest_status_dv)
    retest_status_dv.add('S2:S100')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [10, 15, 15, 20, 12, 20, 30, 12, 12, 12, 12, 10, 12, 15, 15, 35, 12, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_sbom_management_sheet(ws):
    """Create the SBOM Management sheet."""
    ws.title = "SBOM Management"

    headers = [
        "SBOM_ID", "Deliverable_ID", "SBOM_Format", "SBOM_Date",
        "Total_Components", "Direct_Dependencies", "Transitive_Dependencies",
        "Known_Vulnerabilities", "Critical_Vulns", "High_Vulns",
        "License_Issues", "Outdated_Components", "Review_Status",
        "Reviewed_By", "Review_Date", "SBOM_Reference", "Action_Plan"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    format_dv = DataValidation(type="list", formula1='"CycloneDX,SPDX,Spreadsheet,Other"')
    ws.add_data_validation(format_dv)
    format_dv.add('C2:C100')

    status_dv = DataValidation(type="list", formula1='"Pending,Reviewed,Accepted,Rejected"')
    ws.add_data_validation(status_dv)
    status_dv.add('M2:M100')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 12, 12, 15, 18, 20, 18, 12, 12, 12, 18, 12, 20, 12, 35, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_acceptance_signoff_sheet(ws):
    """Create the Acceptance Sign-off sheet."""
    ws.title = "Acceptance Sign-off"

    headers = [
        "Acceptance_ID", "Deliverable_ID", "Criteria_Category", "Acceptance_Criteria",
        "Status", "Evidence_Reference", "Verified_By", "Verification_Date",
        "Waiver_Reason", "Waiver_Approver"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Standard acceptance criteria
    criteria = [
        ("Security", "Code review completed with no unresolved Critical/High findings"),
        ("Security", "SAST scan completed with no unresolved Critical/High findings"),
        ("Security", "SCA scan completed with no Critical/High vulnerable dependencies"),
        ("Security", "DAST scan completed (for web applications)"),
        ("Security", "Penetration test passed (for Critical projects)"),
        ("Security", "SBOM received and reviewed"),
        ("Security", "No secrets detected in codebase"),
        ("Documentation", "Security documentation complete"),
        ("Compliance", "Vulnerability remediation SLAs met"),
        ("Compliance", "Security training completed by all developers"),
    ]

    row = 2
    for category, criterion in criteria:
        ws.cell(row=row, column=1, value=f"ACC-{row-1:04d}").fill = INPUT_FILL
        ws.cell(row=row, column=2, value="[Deliverable_ID]").fill = INPUT_FILL
        ws.cell(row=row, column=3, value=category)
        ws.cell(row=row, column=4, value=criterion)
        for col in range(5, 11):
            ws.cell(row=row, column=col).fill = INPUT_FILL
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Data validation
    category_dv = DataValidation(type="list", formula1='"Functional,Security,Performance,Documentation,Compliance"')
    ws.add_data_validation(category_dv)
    category_dv.add('C2:C100')

    status_dv = DataValidation(type="list", formula1='"Met,Not Met,Waived,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('E2:E100')

    # Column widths
    widths = [12, 15, 18, 55, 10, 35, 20, 15, 35, 25]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_deliverable_inventory_sheet(wb.create_sheet())
    create_code_review_sheet(wb.create_sheet())
    create_security_testing_sheet(wb.create_sheet())
    create_sbom_management_sheet(wb.create_sheet())
    create_acceptance_signoff_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.8.30 Outsourced Development
# =============================================================================
