#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.30.1 - Vendor Assessment and Registry
================================================================================

ISO/IEC 27001:2022 Control A.8.30: Outsourced Development
Assessment Domain 1 of 4: Vendor Assessment and Registry

This script generates a comprehensive Excel assessment workbook for managing
vendor security assessments and the approved vendor registry for outsourced
development engagements.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.30.1"
WORKBOOK_NAME = "Vendor Assessment and Registry"
CONTROL_ID = "A.8.30"
CONTROL_NAME = "Outsourced Development"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.8.30.1 - Vendor Assessment and Registry"],
        [""],
        ["PURPOSE"],
        ["This workbook tracks vendor security assessments and maintains the approved vendor registry"],
        ["for outsourced development engagements per ISMS-POL-A.8.30."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Vendor Registry - Approved vendor list with status and classifications"],
        ["3. Security Assessment - Detailed assessment records per vendor"],
        ["4. Due Diligence Checklist - Verification items for vendor onboarding"],
        ["5. Environment Security - Development environment security attestations"],
        [""],
        ["WORKFLOW"],
        ["1. Identify vendor for engagement"],
        ["2. Classify vendor risk tier (Critical/High/Standard)"],
        ["3. Send security questionnaire"],
        ["4. Complete due diligence checks"],
        ["5. Assess development environment security"],
        ["6. Assign risk rating and recommendation"],
        ["7. Obtain approval from appropriate authority"],
        ["8. Add to approved vendor registry"],
        ["9. Schedule annual reassessment"],
        [""],
        ["DATA VALIDATION"],
        ["- Yellow cells are input fields"],
        ["- Dropdown lists enforce valid values"],
        ["- Vendor_ID must be unique (VND-XXXX format)"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "WORKFLOW", "DATA VALIDATION"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 80


def create_vendor_registry_sheet(ws):
    """Create the Vendor Registry sheet."""
    ws.title = "Vendor Registry"

    headers = [
        "Vendor_ID", "Vendor_Name", "Registry_Status", "Risk_Tier",
        "Initial_Assessment_Date", "Last_Assessment_Date", "Next_Assessment_Due",
        "ISO_27001_Certified", "SOC2_Type2", "Primary_Contact",
        "Approved_Project_Types", "Approved_By", "Notes"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Add data validation dropdowns
    status_dv = DataValidation(type="list", formula1='"Approved,Pending,Suspended,Removed"')
    ws.add_data_validation(status_dv)
    status_dv.add(f'C2:C100')

    tier_dv = DataValidation(type="list", formula1='"Critical,High,Standard"')
    ws.add_data_validation(tier_dv)
    tier_dv.add(f'D2:D100')

    cert_dv = DataValidation(type="list", formula1='"Yes,No,In Progress"')
    ws.add_data_validation(cert_dv)
    cert_dv.add(f'H2:H100')
    cert_dv.add(f'I2:I100')

    # Format input rows
    for row in range(2, 21):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Set column widths
    widths = [12, 30, 15, 12, 18, 18, 18, 15, 12, 25, 25, 20, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_security_assessment_sheet(ws):
    """Create the Security Assessment sheet."""
    ws.title = "Security Assessment"

    headers = [
        "Assessment_ID", "Vendor_ID", "Assessment_Date", "Assessment_Type",
        "Assessor", "Security_Certification", "Cert_Expiry_Date",
        "SDLC_Maturity", "Security_Incident_History", "SAST_DAST_Tooling",
        "Personnel_Screening", "Dev_Environment_Security", "Overall_Risk_Rating",
        "Recommendation", "Conditions", "Evidence_Location"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Initial,Annual,Triggered"')
    ws.add_data_validation(type_dv)
    type_dv.add('D2:D100')

    cert_dv = DataValidation(type="list", formula1='"ISO 27001,SOC 2,Both,None"')
    ws.add_data_validation(cert_dv)
    cert_dv.add('F2:F100')

    maturity_dv = DataValidation(type="list", formula1='"Mature,Developing,Basic,Unknown"')
    ws.add_data_validation(maturity_dv)
    maturity_dv.add('H2:H100')

    incident_dv = DataValidation(type="list", formula1='"None,Minor,Major"')
    ws.add_data_validation(incident_dv)
    incident_dv.add('I2:I100')

    screening_dv = DataValidation(type="list", formula1='"Verified,Attested,Unknown"')
    ws.add_data_validation(screening_dv)
    screening_dv.add('K2:K100')

    env_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(env_dv)
    env_dv.add('L2:L100')

    risk_dv = DataValidation(type="list", formula1='"Low,Medium,High,Critical"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('M2:M100')

    rec_dv = DataValidation(type="list", formula1='"Approve,Conditional,Reject"')
    ws.add_data_validation(rec_dv)
    rec_dv.add('N2:N100')

    # Format input rows
    for row in range(2, 21):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [15, 12, 15, 15, 20, 18, 15, 15, 20, 25, 18, 20, 18, 15, 30, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_due_diligence_sheet(ws):
    """Create the Due Diligence Checklist sheet."""
    ws.title = "Due Diligence"

    headers = [
        "Vendor_ID", "Check_Category", "Check_Item", "Status",
        "Evidence_Type", "Evidence_Reference", "Verified_By",
        "Verified_Date", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Standard checklist items
    checklist_items = [
        ("Security Certification", "Valid ISO 27001 or SOC 2 Type II certification"),
        ("Security Certification", "Certification scope covers development services"),
        ("Secure Development", "Documented SDLC methodology"),
        ("Secure Development", "Security gates defined in SDLC"),
        ("Secure Development", "SAST/DAST/SCA tools in use"),
        ("Incident History", "No major security incidents in 24 months"),
        ("Incident History", "Incident response process documented"),
        ("Technical Capability", "Code review practices established"),
        ("Technical Capability", "Vulnerability management process"),
        ("Personnel Security", "Background checks for developers"),
        ("Personnel Security", "Security awareness training"),
        ("Subcontractor Mgmt", "Subcontractor security policy"),
        ("Business Continuity", "BCP/DR plans documented"),
        ("Reference Check", "Reference 1 contacted and verified"),
        ("Reference Check", "Reference 2 contacted and verified"),
    ]

    row = 2
    for category, item in checklist_items:
        ws.cell(row=row, column=1, value="[Vendor_ID]").fill = INPUT_FILL
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=item)
        for col in range(4, 10):
            ws.cell(row=row, column=col).fill = INPUT_FILL
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Data validation
    status_dv = DataValidation(type="list", formula1='"Complete,Pending,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('D2:D100')

    evidence_dv = DataValidation(type="list", formula1='"Certificate,Attestation,Document,Interview"')
    ws.add_data_validation(evidence_dv)
    evidence_dv.add('E2:E100')

    # Column widths
    widths = [12, 20, 45, 12, 15, 35, 20, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_environment_security_sheet(ws):
    """Create the Environment Security Assessment sheet."""
    ws.title = "Environment Security"

    headers = [
        "Vendor_ID", "Assessment_Date", "MFA_Enabled", "Network_Isolation",
        "Endpoint_Security", "Code_Repository", "Secret_Scanning",
        "Branch_Protection", "Data_Handling", "Attestation_Received",
        "Attestation_Date", "Compliance_Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('G2:G100')
    yn_dv.add('J2:J100')

    ynp_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(ynp_dv)
    ynp_dv.add('C2:C100')
    ynp_dv.add('H2:H100')

    network_dv = DataValidation(type="list", formula1='"Isolated,Segmented,Shared"')
    ws.add_data_validation(network_dv)
    network_dv.add('D2:D100')

    endpoint_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(endpoint_dv)
    endpoint_dv.add('E2:E100')

    repo_dv = DataValidation(type="list", formula1='"Secure,Partial,Unsecure"')
    ws.add_data_validation(repo_dv)
    repo_dv.add('F2:F100')

    data_dv = DataValidation(type="list", formula1='"No Prod Data,Masked,Raw"')
    ws.add_data_validation(data_dv)
    data_dv.add('I2:I100')

    compliance_dv = DataValidation(type="list", formula1='"Compliant,Conditional,Non-Compliant"')
    ws.add_data_validation(compliance_dv)
    compliance_dv.add('L2:L100')

    # Format input rows
    for row in range(2, 21):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 12, 15, 15, 15, 15, 15, 15, 18, 15, 18]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_vendor_registry_sheet(wb.create_sheet())
    create_security_assessment_sheet(wb.create_sheet())
    create_due_diligence_sheet(wb.create_sheet())
    create_environment_security_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.8.30 Outsourced Development
# =============================================================================
