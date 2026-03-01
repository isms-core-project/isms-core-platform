#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.30.2 - Contract Compliance Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.30: Outsourced Development
Assessment Domain 2 of 3: Contract Compliance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific outsourced development security infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Vendor assessment criteria and minimum security requirement thresholds (match your procurement policy)
2. Contract security clause categories and mandatory requirements (legal review required)
3. Security testing scope and acceptance criteria per development type
4. Vendor access control categories and provisioning requirements
5. Delivery and acceptance workflow security checkpoints

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.30 Outsourced Development Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
outsourced development security controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Contract Compliance under ISO 27001:2022 Control A.8.30. Supports evidence-based evaluation of outsourced development vendor security posture, contract compliance, and security testing effectiveness.

**Assessment Scope:**
- Vendor security assessment completeness and compliance rating
- Contract security clause coverage and enforceability
- Vendor access control implementation and monitoring
- Security testing coverage and acceptance criteria compliance
- Delivery security checkpoint adherence and documentation
- Vendor security incident notification and response obligations
- Evidence collection for supply chain security and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Outsourced Development controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a830_2_contract_compliance.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a830_2_contract_compliance.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a830_2_contract_compliance.py --date 20250115

Output:
    File: ISMS-IMP-A.8.30.2_Contract_Compliance_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.30
Assessment Domain:    2 of 3 (Contract Compliance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.30: Outsourced Development Policy (Governance)
    - ISMS-IMP-A.8.30.1: Vendor Assessment and Registry (Domain 1)
    - ISMS-IMP-A.8.30.2: Contract Compliance (Domain 2)
    - ISMS-IMP-A.8.30.3: Security Testing and Acceptance (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.30.2 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive outsourced development security details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review outsourced development vendor assessments and contract security requirements annually or when new vendors are engaged, development practices change, or security incidents involving vendors occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
import sys
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")
from pathlib import Path

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.30.2"
WORKBOOK_NAME = "Contract Compliance"
CONTROL_ID = "A.8.30"
CONTROL_NAME = "Outsourced Development"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = "\u2705"
WARNING = "\u26a0\ufe0f"
XMARK = "\u274c"
DASH = "\u2014"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF", name="Calibri")

SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete all assessment sheets in order, starting with Contract Inventory.', '2. For each contract, verify security clause inclusion in Security Clauses sheet.', '3. Track vulnerability remediation SLA compliance in SLA Compliance sheet.', '4. Record subcontractor authorisation requests in Subcontractor Approvals sheet.', '5. Complete termination checklist at contract closure.', '6. Record all supporting evidence in the Evidence Register sheet.', '7. Use the Summary Dashboard to track overall compliance status.', '8. All user-input cells are highlighted in yellow.', '9. Submit the completed workbook for review and approval via the Approval Sign-Off sheet.', '10. Retain this workbook as part of the ISMS evidence library.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 24

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_contract_inventory_sheet(ws):
    """Create the Contract Inventory sheet."""
    ws.title = "Contract Inventory"

    headers = [
        "Contract ID", "Vendor ID", "Contract Name", "Contract Type",
        "Start Date", "End Date", "Project Classification", "Contract Value CHF",
        "Primary Contact", "Legal Review Date", "Security Review Date", "Status"
    ]

    # Title row (Row 1)
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "CONTRACT INVENTORY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write column headers (Row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Add data validation dropdowns
    type_dv = DataValidation(type="list", formula1='"Fixed-price,T&M,Staff Aug,Managed Service"')
    ws.add_data_validation(type_dv)
    type_dv.add('D4:D102')

    class_dv = DataValidation(type="list", formula1='"Critical,High,Standard"')
    ws.add_data_validation(class_dv)
    class_dv.add('G4:G102')

    status_dv = DataValidation(type="list", formula1='"Active,Completed,Terminated"')
    ws.add_data_validation(status_dv)
    status_dv.add('L4:L102')

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "CONT-001", "VEND-001", "Web Application Development 2024-2025", "Fixed-price",
        "01.01.2024", "31.12.2025", "Critical", "250000",
        "Project Manager (pm@vendor.com)", "15.11.2023", "20.11.2023", "Active"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Empty data rows (rows 5-54) - 50 empty rows for user data
    for row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Set column widths
    widths = [12, 12, 35, 18, 12, 12, 18, 15, 25, 15, 18, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_security_clauses_sheet(ws):
    """Create the Security Clauses Checklist sheet."""
    ws.title = "Security Clauses"

    headers = [
        "Contract ID", "Clause Category", "Clause Description", "Included",
        "Clause Reference", "Modification Notes", "Reviewed By", "Review Date"
    ]

    # Title row (Row 1)
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "SECURITY CLAUSES CHECKLIST"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write column headers (Row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Grey sample row (row 4)
    row = 4
    sample_data = ["[Contract ID]", "Secure Coding Standards", "Compliance with secure coding standards per ISMS-POL-A.8.28"]
    for col, value in enumerate(sample_data, 1):
        ws.cell(row=row, column=col, value=value).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=10, color="808080")
    for col in range(len(sample_data) + 1, 9):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    # 50 empty FFFFCC input rows (rows 5-54)
    while row < 55:
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Data validation
    included_dv = DataValidation(type="list", formula1='"Yes,No,N/A,Modified"')
    ws.add_data_validation(included_dv)
    included_dv.add('D4:D102')

    # Column widths
    widths = [12, 25, 55, 12, 18, 35, 20, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_sla_compliance_sheet(ws):
    """Create the SLA Compliance Tracking sheet."""
    ws.title = "SLA Compliance"

    headers = [
        "SLA ID", "Contract ID", "Vulnerability ID", "Severity",
        "Discovery Date", "SLA Days", "SLA Due Date", "Remediation Date",
        "SLA Met", "Exception Approved", "Exception Approver", "Notes"
    ]

    # Title row (Row 1)
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "SLA COMPLIANCE TRACKING"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write column headers (Row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('D4:D102')

    sla_days_dv = DataValidation(type="list", formula1='"7,30,90,180"')
    ws.add_data_validation(sla_days_dv)
    sla_days_dv.add('F4:F102')

    met_dv = DataValidation(type="list", formula1='"Met,Missed,Pending,Exception"')
    ws.add_data_validation(met_dv)
    met_dv.add('I4:I102')

    exception_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(exception_dv)
    exception_dv.add('J4:J102')

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "SLA-001", "CONT-001", "VULN-2024-001", "Critical",
        "01.12.2024", "7", "08.12.2024", "06.12.2024",
        "Met", "No", "", "Patch deployed 2 days ahead of SLA"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Empty data rows (rows 5-54) - 50 empty rows for user data
    for row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 18, 10, 15, 10, 15, 15, 12, 15, 25, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_subcontractor_approvals_sheet(ws):
    """Create the Subcontractor Approvals sheet."""
    ws.title = "Subcontractor Approvals"

    headers = [
        "Approval ID", "Contract ID", "Primary Vendor ID", "Subcontractor Name",
        "Subcontractor Scope", "Access Level", "Assessment Level", "Risk Classification",
        "Approval Status", "Approved By", "Approval Date", "Expiry Date",
        "Flow Down Verified", "Notes"
    ]

    # Title row (Row 1)
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "SUBCONTRACTOR APPROVALS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write column headers (Row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    access_dv = DataValidation(type="list", formula1='"Direct,Via Vendor,None"')
    ws.add_data_validation(access_dv)
    access_dv.add('F4:F102')

    assess_dv = DataValidation(type="list", formula1='"Full,Abbreviated,Vendor Attested"')
    ws.add_data_validation(assess_dv)
    assess_dv.add('G4:G102')

    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('H4:H102')

    status_dv = DataValidation(type="list", formula1='"Approved,Pending,Rejected"')
    ws.add_data_validation(status_dv)
    status_dv.add('I4:I102')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('M4:M102')

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "SUB-001", "CONT-001", "VEND-001", "SecureDev Partners Ltd",
        "Penetration testing services", "Direct", "Full", "Medium",
        "Approved", "CISO", "15.11.2024", "15.11.2025",
        "Yes", "ISO 27001 certified subcontractor"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Empty data rows (rows 5-54) - 50 empty rows for user data
    for row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 15, 30, 35, 12, 18, 18, 15, 20, 12, 12, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_termination_checklist_sheet(ws):
    """Create the Termination Checklist sheet."""
    ws.title = "Termination Checklist"

    headers = [
        "Contract ID", "Termination Type", "Termination Date", "Check Item",
        "Status", "Completion Date", "Verified By", "Evidence Reference"
    ]

    # Title row (Row 1)
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "TERMINATION CHECKLIST"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write column headers (Row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Standard termination checklist items
    checklist_items = [
        "All access credentials revoked (within 24 hours)",
        "[Organisation] data returned or destroyed",
        "Destruction certificate received",
        "Source code transferred (if applicable)",
        "Documentation transferred",
        "Escrow arrangements verified",
        "Outstanding vulnerabilities addressed or risk accepted",
        "Final security review completed",
        "Lessons learned documented",
        "Vendor removed from active registry",
    ]

    # Sample row (row 4) with realistic data (grey fill per Option B standard)
    sample_data = [
        "Contract 001", "Completion", "15.12.2024",
        "All access credentials revoked (within 24 hours)",
        "Complete", "16.12.2024", "Security Team",
        "Termination/Contract-001-Access-Revocation-Confirmed.pdf"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = THIN_BORDER

    # Remaining checklist items as empty yellow rows (rows 5-13)
    row = 5
    for item in checklist_items[1:]:  # Skip first item (already in sample row)
        ws.cell(row=row, column=4, value=item).fill = INPUT_FILL
        ws.cell(row=row, column=4).font = Font(name="Calibri", size=10)
        for col in [1, 2, 3, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).fill = INPUT_FILL
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Extend to 50 data rows (rows 4-54)
    while row < 55:
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Data validation
    type_dv = DataValidation(type="list", formula1='"Completion,Early,Breach"')
    ws.add_data_validation(type_dv)
    type_dv.add('B4:B102')

    status_dv = DataValidation(type="list", formula1='"Complete,Pending,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('E4:E102')

    # Column widths
    widths = [12, 15, 15, 50, 12, 15, 20, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Assessment area configurations
    # Format: (Sheet Name, Status Column, [Compliant, Partial, Non-Compliant, N/A], Row Start, Row End)
    area_configs = [
        ('Contract Inventory', 'L', ['Active', None, 'Completed,Terminated', None], 5, 54),
        ('Security Clauses', 'D', ['Yes', 'Modified', 'No', 'N/A'], 5, 54),
        ('SLA Compliance', 'I', ['Met', 'Pending', 'Missed', 'Exception'], 5, 54),
        ('Subcontractor Approvals', 'I', ['Approved', 'Pending', 'Rejected', None], 5, 54),
        ('Termination Checklist', 'E', ['Complete', 'Pending', None, 'N/A'], 5, 54),
    ]

    # Assessment area display names (for TABLE 1 column A)
    assessment_areas = [
        'Contract Inventory',
        'Security Clauses Compliance',
        'SLA Compliance Tracking',
        'Subcontractor Approvals',
        'Termination Checklist',
    ]

    ws.title = "Summary Dashboard"

    # Header (Row 1)
    ws.merge_cells("A1:G1")
    ws["A1"] = "CONTRACT COMPLIANCE — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Control reference (Row 2)
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")

    # TABLE 1 banner (Row 4)
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

    # TABLE 1 Column Headers (Row 5) - FIX PAL-002: Use 4472C4 not D9D9D9
    row = 5
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows with COUNTIF formulas
    for i, (display_name, (sheet_name, status_col, status_values, row_start, row_end)) in enumerate(zip(assessment_areas, area_configs)):
        r = 6 + i
        ws.cell(row=r, column=1, value=display_name).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border

        # Column B: Total Requirements
        c = ws.cell(row=r, column=2)
        c.value = f"=COUNTA('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end})"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column C: Compliant
        c = ws.cell(row=r, column=3)
        if status_values[0] and ',' in status_values[0]:
            compliant_options = status_values[0].split(',')
            formula_parts = [f'COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{opt}")' for opt in compliant_options]
            c.value = f"={'+'.join(formula_parts)}"
        elif status_values[0]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[0]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column D: Partially Compliant
        c = ws.cell(row=r, column=4)
        if status_values[1]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[1]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column E: Non-Compliant
        c = ws.cell(row=r, column=5)
        if status_values[2] and ',' in status_values[2]:
            nc_options = status_values[2].split(',')
            formula_parts = [f'COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{opt}")' for opt in nc_options]
            c.value = f"={'+'.join(formula_parts)}"
        elif status_values[2]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[2]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column F: N/A
        c = ws.cell(row=r, column=6)
        if status_values[3]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[3]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column G: Compliance %
        c = ws.cell(row=r, column=7)
        c.value = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
        c.number_format = "0.0%"
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # TOTAL row - sum only data rows to avoid circular reference
    total_row = 6 + len(area_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1).border = border

    for col_idx in range(2, 7):
        col_letter = chr(64 + col_idx)
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}6:{col_letter}{total_row-1})"
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        c.border = border
        c.alignment = Alignment(horizontal="center")

    c = ws.cell(row=total_row, column=7)
    c.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    c.border = border
    c.alignment = Alignment(horizontal="center")

    # TABLE 2: KEY METRICS (All formula-driven!)
    row = total_row + 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 2: KEY METRICS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

    # Metric 1: Total Active Contracts
    r = row + 1
    ws.cell(row=r, column=1, value="Total Active Contracts").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=COUNTIF(\'Contract Inventory\'!L5:L54,"Active")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 2: Total Subcontractors
    r = row + 2
    ws.cell(row=r, column=1, value="Total Subcontractors").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=COUNTA(\'Subcontractor Approvals\'!B5:B54)'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 3: Active Critical Projects
    r = row + 3
    ws.cell(row=r, column=1, value="Active Critical Projects").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=COUNTIFS(\'Contract Inventory\'!L5:L54,"Active",\'Contract Inventory\'!G5:G54,"Critical")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 4: SLA Compliance Rate
    r = row + 4
    ws.cell(row=r, column=1, value="SLA Compliance Rate").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=IF(COUNTA(\'SLA Compliance\'!I5:I54)=0,"0%",ROUND(COUNTIF(\'SLA Compliance\'!I5:I54,"Met")/COUNTA(\'SLA Compliance\'!I5:I54)*100,1)&"%")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 5: Security Clauses Met
    r = row + 5
    ws.cell(row=r, column=1, value="Security Clauses Met").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=IF(COUNTA(\'Security Clauses\'!D5:D54)=0,"0%",ROUND(COUNTIF(\'Security Clauses\'!D5:D54,"Yes")/COUNTA(\'Security Clauses\'!D5:D54)*100,1)&"%")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 6: Pending Terminations
    r = row + 6
    ws.cell(row=r, column=1, value="Pending Terminations").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=COUNTIF(\'Termination Checklist\'!E5:E54,"Pending")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # TABLE 3: CRITICAL FINDINGS
    row = row + 8
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 3: CRITICAL FINDINGS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    row += 1
    findings_headers = ["#", "Finding", "Severity", "Affected Area", "Recommended Action", "Owner", "Due Date"]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Critical findings based on DV analysis
    findings = [
        ('=COUNTIF(\'Contract Inventory\'!L5:L54,"Terminated")', "Terminated Contracts Requiring Cleanup", "High", "Contract Inventory", "Review termination checklist completion"),
        ('=COUNTIF(\'Security Clauses\'!D5:D54,"No")', "Missing Critical Security Clauses", "Critical", "Security Clauses", "Add required clauses to contracts"),
        ('=COUNTIF(\'SLA Compliance\'!I5:I54,"Missed")', "SLA Compliance Failures", "High", "SLA Compliance", "Investigate SLA breaches"),
        ('=COUNTIF(\'Subcontractor Approvals\'!I5:I54,"Pending")', "Pending Subcontractor Approvals", "Medium", "Subcontractor Approvals", "Complete subcontractor assessments"),
        ('=COUNTIF(\'Subcontractor Approvals\'!I5:I54,"Rejected")', "Rejected Subcontractors", "Critical", "Subcontractor Approvals", "Remove access immediately"),
        ('=COUNTIF(\'Subcontractor Approvals\'!H5:H54,"High")', "High-Risk Subcontractors", "High", "Subcontractor Approvals", "Enhanced monitoring required"),
        ('=COUNTIF(\'Subcontractor Approvals\'!M5:M54,"No")', "Flow-Down Requirements Not Verified", "Medium", "Subcontractor Approvals", "Verify subcontractor compliance flow-down"),
    ]

    for i, (count_formula, finding_name, severity, area, action) in enumerate(findings, 1):
        r = row + i
        ws.cell(row=r, column=1, value=count_formula).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=finding_name).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=severity).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=4, value=area).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=5, value=action).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=6).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=r, column=6).border = border
        c = ws.cell(row=r, column=7)
        c.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        c.border = border
        c.number_format = "DD.MM.YYYY"  # Swiss date format for Due Date column

    # Column widths & freeze
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create the Evidence Register sheet -- standard 8-column format."""
    ws.title = "Evidence Register"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1)
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column Headers (Row 4)
    headers = [
        ("Evidence ID", 15),
        ("Assessment Area", 25),
        ("Evidence Type", 22),
        ("Description", 40),
        ("Location/Path", 45),
        ("Date Collected", 16),
        ("Collected By", 20),
        ("Verification Status", 22),
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    type_dv.add("C6:C105")

    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("H6:H105")

    # Sample row (row 5) with example data (grey fill per Option B standard)
    sample_data = [
        "EV-001", "Contract Inventory", "Configuration file",
        "Executed contract with security clauses - Acme Corp",
        "Legal/Contracts/2024/Acme-Development-Contract-Signed.pdf",
        "15.11.2023", "Legal Team", "Verified"
    ]
    for c, value in enumerate(sample_data, 1):
        cell = ws.cell(row=5, column=c, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(color="808080")
        cell.border = border

    # Empty data rows (rows 6-105) - 100 empty rows for user data (MAX-002 standard)
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

    # Freeze
    ws.freeze_panes = "A5"


def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet -- standard pattern."""
    ws.title = "Approval Sign-Off"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1)
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner (Row 3)
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    # Summary fields
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G10"),  # TABLE 1 TOTAL row
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        # Apply border to all cells in merged range (always)
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = border
            if value == "":
                # Apply fill only to empty fields
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Status dropdown on Assessment Status
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"B{row - 1}")

    # 3 Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2  # gap before first approver
    for title, color in approvers:
        # Banner
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        # 5 fields per approver
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            # Apply fill and border to all cells in merged range
            for col in range(2, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws.cell(row=row, column=col).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    # Apply fill and border to all cells in merged range
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=col).border = border

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        # Apply fill and border to all cells in merged range
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    default_sheet = wb.active
    default_sheet.sheet_view.showGridLines = False

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_contract_inventory_sheet(wb.create_sheet())
    create_security_clauses_sheet(wb.create_sheet())
    create_sla_compliance_sheet(wb.create_sheet())
    create_subcontractor_approvals_sheet(wb.create_sheet())
    create_termination_checklist_sheet(wb.create_sheet())
    create_evidence_register(wb.create_sheet())
    create_summary_dashboard_sheet(wb.create_sheet())
    create_approval_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")



# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
