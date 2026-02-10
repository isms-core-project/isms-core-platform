#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.30.4 - Monitoring and Exceptions Dashboard
================================================================================

ISO/IEC 27001:2022 Control A.8.30: Outsourced Development
Assessment Domain 4 of 4: Monitoring and Exceptions Dashboard

This script generates a comprehensive Excel dashboard workbook that consolidates
outsourced development security status, tracks exceptions, and provides executive
reporting on vendor security compliance.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.30.4"
WORKBOOK_NAME = "Monitoring and Exceptions Dashboard"
CONTROL_ID = "A.8.30"
CONTROL_NAME = "Outsourced Development"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Unicode symbols
CHECK = "\u2705"
WARNING = "\u26a0\ufe0f"
XMARK = "\u274c"
DASH = "\u2014"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions and Legend sheet."""
    ws.title = "Instructions"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1) -- two-line merged header
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Document Information (Row 3+)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID:", DOCUMENT_ID),
        ("Assessment Area:", WORKBOOK_NAME),
        ("Related Policy:", "ISMS-POL-A.8.30"),
        ("Version:", "1.0"),
        ("Assessment Date:", ""),
        ("Completed By:", ""),
        ("Organisation:", ""),
        ("Review Cycle:", "Annually"),
    ]

    for i, (label, value) in enumerate(doc_info):
        row = 4 + i
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border

    # Instructions Section
    ws["A13"] = "Instructions"
    ws["A13"].font = Font(bold=True, size=12)

    instructions = [
        "1. Review the Executive Dashboard for consolidated compliance status.",
        "2. Monitor vendor performance scores and trends in Vendor Performance.",
        "3. Track all security exceptions in the Exception Register.",
        "4. Record ongoing monitoring activities in the Monitoring Log.",
        "5. Log security incidents related to outsourced development in the Incident Log.",
        "6. Review the Compliance Score calculation for overall scoring methodology.",
        "7. Record all supporting evidence in the Evidence Register sheet.",
        "8. All user-input cells are highlighted in yellow.",
        "9. Submit the completed workbook for review and approval via the Approval Sign-Off sheet.",
        "10. Retain this workbook as part of the ISMS evidence library.",
    ]

    for i, text in enumerate(instructions):
        ws[f"A{14 + i}"] = text

    # Acceptable Evidence
    ws["A25"] = "ACCEPTABLE EVIDENCE (examples)"
    ws["A25"].font = Font(bold=True, size=12)

    evidence_items = [
        "Vendor performance scorecards and trend reports",
        "Security exception approval and tracking records",
        "Monitoring activity logs and compliance reports",
        "Security incident reports and remediation evidence",
        "Compliance score calculation worksheets",
        "Executive dashboard snapshots and board reports",
    ]

    for i, item in enumerate(evidence_items):
        ws[f"A{26 + i}"] = f"{DASH} {item}"

    # Status Legend (Table Format)
    legend_row = 33
    ws[f"A{legend_row}"] = "Status Legend"
    ws[f"A{legend_row}"].font = Font(bold=True, size=12)

    legend_headers = ["Symbol", "Status", "Description"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=legend_row + 1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border

    legend_items = [
        (CHECK, "Compliant", "Requirement fully met with evidence"),
        (WARNING, "Partial", "Partially implemented, gaps identified"),
        (XMARK, "Non-Compliant", "Requirement not met, remediation needed"),
        (DASH, "N/A", "Not applicable to this organisation"),
    ]

    for i, (symbol, status, desc) in enumerate(legend_items):
        r = legend_row + 2 + i
        ws.cell(row=r, column=1, value=symbol).border = border
        ws.cell(row=r, column=2, value=status).border = border
        ws.cell(row=r, column=3, value=desc).border = border

    # Column Widths & Freeze
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


def create_executive_dashboard_sheet(ws):
    """Create the Executive Dashboard sheet."""
    ws.title = "Executive Dashboard"

    # Title row (A1) - merged, 003366 fill, white font 14pt bold
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="EXECUTIVE DASHBOARD")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Report date
    ws.cell(row=2, column=1, value=f"Report Date: {GENERATED_DATE}")
    ws.cell(row=2, column=1).font = Font(italic=True)

    # Metrics headers
    col_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    col_header_font = Font(bold=True, size=11, color="000000")
    headers = ["Metric", "Description", "Target", "Current", "Trend"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Dashboard metrics
    metrics = [
        ("Approved Vendors", "Vendors in approved registry", "N/A", "[Count]"),
        ("Active Contracts", "Contracts currently active", "N/A", "[Count]"),
        ("Pending Assessments", "Vendors awaiting assessment", "0", "[Count]"),
        ("Overdue Reassessments", "Annual reassessments overdue", "0", "[Count]"),
        ("Contract Clause Compliance", "Contracts with all security clauses", "100%", "[%]"),
        ("SLA Compliance (Critical)", "Critical vulns fixed within SLA", ">=95%", "[%]"),
        ("SLA Compliance (High)", "High vulns fixed within SLA", ">=90%", "[%]"),
        ("Security Testing Completion", "Deliverables with complete testing", "100%", "[%]"),
        ("SBOM Compliance", "Deliverables with SBOM received", "100%", "[%]"),
        ("Active Exceptions", "Open security exceptions", "[Track]", "[Count]"),
        ("Overdue Exceptions", "Exceptions past expiry", "0", "[Count]"),
        ("Overall Compliance Score", "Weighted compliance score", ">=90%", "[%]"),
    ]

    row = 5
    for metric, desc, target, current in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=3, value=target).border = THIN_BORDER
        cell_current = ws.cell(row=row, column=4, value=current)
        cell_current.fill = INPUT_FILL
        cell_current.border = THIN_BORDER
        cell_trend = ws.cell(row=row, column=5)
        cell_trend.fill = INPUT_FILL
        cell_trend.border = THIN_BORDER
        row += 1

    # Trend validation
    trend_dv = DataValidation(type="list", formula1='"↑,↓,→"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('E5:E20')

    # Column widths
    widths = [30, 40, 15, 15, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_vendor_performance_sheet(ws):
    """Create the Vendor Performance sheet."""
    ws.title = "Vendor Performance"

    # Title row (A1) - merged across all columns
    last_col = get_column_letter(15)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value="VENDOR PERFORMANCE")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Column headers in row 3 (gray fill, dark bold font)
    col_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    col_header_font = Font(bold=True, size=11, color="000000")

    headers = [
        "Vendor_ID", "Vendor_Name", "Risk_Tier", "Active_Contracts",
        "Total_Deliverables", "Security_Findings_Total", "Critical_Findings",
        "High_Findings", "SLA_Compliance_Rate", "Avg_Remediation_Days",
        "Security_Incidents", "Last_Assessment_Date", "Performance_Score",
        "Performance_Trend", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (shifted down by 2)
    tier_dv = DataValidation(type="list", formula1='"Critical,High,Standard"')
    ws.add_data_validation(tier_dv)
    tier_dv.add('C4:C102')

    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('N4:N102')

    # Format input rows (shifted down by 2)
    for row in range(4, 23):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 25, 12, 15, 15, 18, 15, 12, 18, 18, 15, 18, 15, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_exception_register_sheet(ws):
    """Create the Exception Register sheet."""
    ws.title = "Exception Register"

    # Title row (A1) - merged across all columns
    last_col = get_column_letter(19)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value="EXCEPTION REGISTER")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Column headers in row 3 (gray fill, dark bold font)
    col_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    col_header_font = Font(bold=True, size=11, color="000000")

    headers = [
        "Exception_ID", "Exception_Type", "Related_Entity", "Requirement_Reference",
        "Exception_Description", "Risk_Level", "Business_Justification",
        "Compensating_Controls", "Requested_By", "Request_Date", "Approved_By",
        "Approval_Date", "Expiry_Date", "Status", "Renewal_Count",
        "Last_Review_Date", "Next_Review_Date", "Closure_Date", "Closure_Reason"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (shifted down by 2)
    type_dv = DataValidation(type="list", formula1='"Vendor Assessment,Contract Clause,SLA,Testing,Training"')
    ws.add_data_validation(type_dv)
    type_dv.add('B4:B102')

    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('F4:F102')

    status_dv = DataValidation(type="list", formula1='"Pending,Approved,Rejected,Expired,Renewed"')
    ws.add_data_validation(status_dv)
    status_dv.add('N4:N102')

    closure_dv = DataValidation(type="list", formula1='"Remediated,Risk Accepted,Terminated,N/A"')
    ws.add_data_validation(closure_dv)
    closure_dv.add('S4:S102')

    # Format input rows (shifted down by 2)
    for row in range(4, 33):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 18, 15, 20, 35, 12, 35, 35, 15, 12, 20, 12, 12, 12, 12, 15, 15, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_monitoring_log_sheet(ws):
    """Create the Monitoring Log sheet."""
    ws.title = "Monitoring Log"

    # Title row (A1) - merged across all columns
    last_col = get_column_letter(13)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value="MONITORING LOG")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Column headers in row 3 (gray fill, dark bold font)
    col_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    col_header_font = Font(bold=True, size=11, color="000000")

    headers = [
        "Log_ID", "Log_Date", "Vendor_ID", "Contract_ID", "Activity_Type",
        "Activity_Description", "Participants", "Findings", "Actions_Required",
        "Action_Owner", "Action_Due_Date", "Action_Status", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (shifted down by 2)
    type_dv = DataValidation(type="list", formula1='"Status Meeting,Security Review,Audit,Incident Review,Ad-hoc"')
    ws.add_data_validation(type_dv)
    type_dv.add('E4:E102')

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Complete,Overdue"')
    ws.add_data_validation(status_dv)
    status_dv.add('L4:L102')

    # Format input rows (shifted down by 2)
    for row in range(4, 53):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 12, 12, 18, 40, 25, 35, 35, 20, 15, 15, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_incident_log_sheet(ws):
    """Create the Incident Log sheet."""
    ws.title = "Incident Log"

    # Title row (A1) - merged across all columns
    last_col = get_column_letter(17)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1, value="INCIDENT LOG")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Column headers in row 3 (gray fill, dark bold font)
    col_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    col_header_font = Font(bold=True, size=11, color="000000")

    headers = [
        "Incident_ID", "Incident_Date", "Vendor_ID", "Contract_ID", "Incident_Type",
        "Severity", "Description", "Root_Cause", "Impact_Assessment",
        "Notification_Date", "Notification_SLA_Met", "Remediation_Actions",
        "Remediation_Date", "Lessons_Learned", "Status", "Closed_Date", "Contract_Impact"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations (shifted down by 2)
    type_dv = DataValidation(type="list", formula1='"Data Breach,Vulnerability Exploited,Unauthorised Access,Policy Violation,Other"')
    ws.add_data_validation(type_dv)
    type_dv.add('E4:E102')

    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('F4:F102')

    sla_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(sla_dv)
    sla_dv.add('K4:K102')

    status_dv = DataValidation(type="list", formula1='"Open,Investigating,Remediated,Closed"')
    ws.add_data_validation(status_dv)
    status_dv.add('O4:O102')

    impact_dv = DataValidation(type="list", formula1='"None,Warning,Review,Suspension,Termination"')
    ws.add_data_validation(impact_dv)
    impact_dv.add('Q4:Q102')

    # Format input rows (shifted down by 2)
    for row in range(4, 33):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 12, 12, 20, 10, 40, 35, 35, 15, 15, 40, 15, 35, 12, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_compliance_score_sheet(ws):
    """Create the Compliance Score Calculation sheet."""
    ws.title = "Compliance Score"

    # Title row (A1) - merged, 003366 fill, white font 14pt bold
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="COMPLIANCE SCORE")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Column headers in row 3 (gray fill, dark bold font)
    col_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    col_header_font = Font(bold=True, size=11, color="000000")
    headers = ["Component", "Weight", "Scoring Criteria", "Data Source", "Raw Score", "Weighted Score"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Score components
    components = [
        ("Vendor Assessment", "20%", "% vendors with current assessment", "Workbook 1"),
        ("Contract Compliance", "25%", "% contracts with all security clauses", "Workbook 2"),
        ("SLA Compliance", "25%", "Weighted avg: Critical (40%) + High (60%)", "Workbook 2"),
        ("Security Testing", "20%", "% deliverables with complete testing", "Workbook 3"),
        ("Exception Management", "10%", "100% - (Overdue exceptions x 10%)", "Sheet 3"),
    ]

    row = 4
    for component, weight, criteria, source in components:
        ws.cell(row=row, column=1, value=component).border = THIN_BORDER
        ws.cell(row=row, column=2, value=weight).border = THIN_BORDER
        ws.cell(row=row, column=3, value=criteria).border = THIN_BORDER
        ws.cell(row=row, column=4, value=source).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = INPUT_FILL
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=6).fill = LOCKED_FILL
        ws.cell(row=row, column=6).border = THIN_BORDER
        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="OVERALL COMPLIANCE SCORE").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=6).fill = GREEN_FILL
    ws.cell(row=row, column=6).border = THIN_BORDER
    ws.cell(row=row, column=6).font = Font(bold=True)

    # Score interpretation
    row += 3
    ws.cell(row=row, column=1, value="Score Interpretation:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=">= 90%: Compliant")
    ws.cell(row=row, column=2).fill = GREEN_FILL
    row += 1
    ws.cell(row=row, column=1, value="70-89%: Needs Improvement")
    ws.cell(row=row, column=2).fill = YELLOW_FILL
    row += 1
    ws.cell(row=row, column=1, value="< 70%: Non-Compliant (Escalation Required)")
    ws.cell(row=row, column=2).fill = RED_FILL

    # Formula note
    row += 2
    ws.cell(row=row, column=1, value="Formula: (Vendor x 0.20) + (Contract x 0.25) + (SLA x 0.25) + (Testing x 0.20) + (Exception x 0.10)")
    ws.cell(row=row, column=1).font = Font(italic=True)

    # Column widths
    widths = [25, 10, 45, 15, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_register_sheet(ws):
    """Create standard Evidence Register (8 columns, 100 rows)."""
    ws.title = "Evidence Register"

    # Header
    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value="EVIDENCE REGISTER")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    cell = ws.cell(row=2, column=1,
                   value="Record all evidence collected during the assessment. "
                         "Each row represents one piece of evidence.")
    cell.font = Font(name="Calibri", size=10, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column headers row 4
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 30

    # 100 data rows (5-104)
    for i in range(1, 101):
        row = i + 4
        # Evidence ID (gray font, no yellow fill)
        cell = ws.cell(row=row, column=1, value=f"EV-{i:03d}")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = THIN_BORDER

        # Cols B-H: yellow fill + border
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor="FFFFCC")
            cell.border = THIN_BORDER

    # Dropdowns
    ev_types = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Log extract,Policy document,'
                 'Training record,Audit report,Risk assessment,Interview notes,'
                 'Test results,Other"',
        allow_blank=True
    )
    ev_types.prompt = "Select evidence type"
    ws.add_data_validation(ev_types)
    ev_types.add("C5:C104")

    verify_status = DataValidation(
        type="list",
        formula1='"Verified,Pending Verification,Insufficient,Not Reviewed"',
        allow_blank=True
    )
    verify_status.prompt = "Select verification status"
    ws.add_data_validation(verify_status)
    verify_status.add("H5:H104")

    # Column widths
    widths = {"A": 15, "B": 25, "C": 22, "D": 40, "E": 45, "F": 16, "G": 20, "H": 22}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A5"


def create_approval_signoff_sheet(ws):
    """Create standard Approval Sign-Off sheet."""
    ws.title = "Approval Sign-Off"

    # Header
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value="ASSESSMENT APPROVAL AND SIGN-OFF")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="ASSESSMENT SUMMARY")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Summary fields
    summary_fields = [
        ("Document:", DOCUMENT_ID),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", ""),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    for i, (label, value) in enumerate(summary_fields):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws.merge_cells(f"B{r}:E{r}")
        cell = ws.cell(row=r, column=2, value=value)
        cell.border = THIN_BORDER
        if label in ("Assessment Status:", "Overall Compliance Rating:"):
            cell.fill = PatternFill("solid", fgColor="FFFFCC")

    # Assessment Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"B{row + 4}")

    row = row + 1 + len(summary_fields) + 1

    # Helper for approver sections
    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        fields = ["Name:", "Title:", "Date:", "Signature:", "Comments:"]
        for idx, field in enumerate(fields):
            r = start_row + 1 + idx
            ws.cell(row=r, column=1, value=field).font = Font(name="Calibri", size=10, bold=True)
            ws.merge_cells(f"B{r}:E{r}")
            c = ws.cell(row=r, column=2)
            c.fill = PatternFill("solid", fgColor="FFFFCC")
            c.border = THIN_BORDER
        return start_row + 1 + len(fields) + 1

    row = _approver_section(row, f"COMPLETED BY {DASH} Assessment Lead", "4472C4")
    row = _approver_section(row, f"REVIEWED BY {DASH} Security Manager", "4472C4")
    row = _approver_section(row, f"APPROVED BY {DASH} CISO", "003366")

    # FINAL DECISION
    ws.cell(row=row, column=1, value="FINAL DECISION:").font = Font(name="Calibri", size=11, bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    cell = ws.cell(row=row, column=2)
    cell.fill = PatternFill("solid", fgColor="FFFFCC")
    cell.border = THIN_BORDER

    final_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(final_dv)
    final_dv.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="NEXT REVIEW DETAILS")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="left", vertical="center")

    review_fields = ["Next Review Date:", "Review Frequency:", "Scheduled Reviewer:"]
    for i, field in enumerate(review_fields):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=field).font = Font(name="Calibri", size=10, bold=True)
        ws.merge_cells(f"B{r}:E{r}")
        cell = ws.cell(row=r, column=2)
        cell.fill = PatternFill("solid", fgColor="FFFFCC")
        cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    ws.freeze_panes = "A3"


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_executive_dashboard_sheet(wb.create_sheet())
    create_vendor_performance_sheet(wb.create_sheet())
    create_exception_register_sheet(wb.create_sheet())
    create_monitoring_log_sheet(wb.create_sheet())
    create_incident_log_sheet(wb.create_sheet())
    create_compliance_score_sheet(wb.create_sheet())
    create_evidence_register_sheet(wb.create_sheet())
    create_approval_signoff_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-02-10
# QA_STATUS: PASSED - STANDARDISATION COMPLETE
# QA_TOOL: Claude Code
# CHANGES: Unicode symbols, freeze panes, Evidence Register, Approval Sign-Off
# =============================================================================
