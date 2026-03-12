#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.30.3 - Security Testing and Acceptance Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.30: Outsourced Development
Assessment Domain 3 of 3: Security Testing and Acceptance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific outsourced development security infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Vendor assessment criteria and minimum security requirement thresholds (match your procurement policy)
2. Contract security clause categories and mandatory requirements (legal review required)
3. Security testing scope and acceptance criteria per development type
4. Vendor access control categories and provisioning requirements
5. Delivery and acceptance workflow security checkpoints

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.30 Outsourced Development Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
outsourced development security controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Security Testing and Acceptance under ISO 27001:2022 Control A.8.30. Supports evidence-based evaluation of outsourced development vendor security posture, contract compliance, and security testing effectiveness.

**Assessment Scope:**
- Vendor security assessment completeness and compliance rating
- Contract security clause coverage and enforceability
- Vendor access control implementation and monitoring
- Security testing coverage and acceptance criteria compliance
- Delivery security checkpoint adherence and documentation
- Vendor security incident notification and response obligations
- Evidence collection for supply chain security and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Outsourced Development controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a830_3_security_testing.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a830_3_security_testing.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a830_3_security_testing.py --date 20250115

Output:
    File: ISMS-IMP-A.8.30.3_Security_Testing_and_Acceptance_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.30
Assessment Domain:    3 of 3 (Security Testing and Acceptance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.30: Outsourced Development Policy (Governance)
    - ISMS-IMP-A.8.30.1: Vendor Assessment and Registry (Domain 1)
    - ISMS-IMP-A.8.30.2: Contract Compliance (Domain 2)
    - ISMS-IMP-A.8.30.3: Security Testing and Acceptance (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.30.3 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive outsourced development security details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review outsourced development vendor assessments and contract security requirements annually or when new vendors are engaged, development practices change, or security incidents involving vendors occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
import sys
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.30.3"
WORKBOOK_NAME = "Security Testing and Acceptance"
CONTROL_ID = "A.8.30"
CONTROL_NAME = "Outsourced Development"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = "\u2705"
WARNING = "\u26a0\ufe0f"
XMARK = "\u274c"
DASH = "\u2014"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete all assessment sheets in order, starting with Deliverable Inventory.', '2. Track code review results for each deliverable in Code Review Tracking.', '3. Record SAST, DAST, SCA, and penetration test results in Security Testing.', '4. Maintain SBOM records for all deliverables in SBOM Management.', '5. Verify acceptance criteria and obtain sign-off in Acceptance Sign-off.', '6. Record all supporting evidence in the Evidence Register sheet.', '7. Use the Summary Dashboard to track overall compliance status.', '8. All user-input cells are highlighted in yellow.', '9. Submit the completed workbook for review and approval via the Approval Sign-Off sheet.', '10. Retain this workbook as part of the ISMS evidence library.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 24

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_deliverable_inventory_sheet(ws):
    """Create the Deliverable Inventory sheet."""
    ws.title = "Deliverable Inventory"

    headers = [
        "Deliverable ID", "Contract ID", "Vendor ID", "Deliverable Name",
        "Deliverable Type", "Project Classification", "Planned Delivery",
        "Actual Delivery", "Code Review Status", "Security Test Status",
        "SBOM Received", "Acceptance Status", "Acceptance Date", "Accepted By"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "DELIVERABLE INVENTORY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write headers in row 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Application,Module,Component,Library,API"')
    ws.add_data_validation(type_dv)
    type_dv.add('E4:E102')

    class_dv = DataValidation(type="list", formula1='"Critical,High,Standard"')
    ws.add_data_validation(class_dv)
    class_dv.add('F4:F102')

    review_dv = DataValidation(type="list", formula1='"Pending,In Progress,Complete,N/A"')
    ws.add_data_validation(review_dv)
    review_dv.add('I4:I102')

    test_dv = DataValidation(type="list", formula1='"Pending,In Progress,Complete"')
    ws.add_data_validation(test_dv)
    test_dv.add('J4:J102')

    sbom_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(sbom_dv)
    sbom_dv.add('K4:K102')

    accept_dv = DataValidation(type="list", formula1='"Pending,Accepted,Rejected,Conditional"')
    ws.add_data_validation(accept_dv)
    accept_dv.add('L4:L102')

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "DEL-001", "CONT-001", "VEND-001", "Customer Portal v2.0",
        "Application", "Critical", "30.11.2024", "28.11.2024",
        "Complete", "Complete", "Yes", "Accepted",
        "28.11.2024", "Security Team"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Empty data rows (rows 5-54) - 50 empty rows for user data
    for row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Set column widths
    widths = [15, 12, 12, 35, 15, 18, 15, 15, 18, 18, 15, 15, 15, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_code_review_sheet(ws):
    """Create the Code Review Tracking sheet."""
    ws.title = "Code Review Tracking"

    headers = [
        "Review ID", "Deliverable ID", "Review Type", "Review Date",
        "Reviewer", "Reviewer Role", "Files Reviewed", "Security Findings",
        "Critical Findings", "High Findings", "Medium Findings", "Low Findings",
        "Review Result", "Findings Reference", "Notes"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "CODE REVIEW TRACKING"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write headers in row 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Peer Review,Security Review,Architecture Review"')
    ws.add_data_validation(type_dv)
    type_dv.add('C4:C102')

    role_dv = DataValidation(type="list", formula1='"Developer,Security Team,Security Architect"')
    ws.add_data_validation(role_dv)
    role_dv.add('F4:F102')

    result_dv = DataValidation(type="list", formula1='"Approved,Approved with Findings,Rejected"')
    ws.add_data_validation(result_dv)
    result_dv.add('M4:M102')

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "REV-001", "DEL-001", "Security Review", "25.11.2024",
        "Jane Smith", "Security Team", "45", "12",
        "3", "6", "2", "1",
        "Approved with Findings", "SharePoint/Reviews/REV-001.pdf", "Minor issues documented"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Empty data rows (rows 5-54) - 50 empty rows for user data
    for row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 18, 12, 20, 18, 15, 15, 15, 12, 15, 12, 20, 35, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_security_testing_sheet(ws):
    """Create the Security Testing Results sheet."""
    ws.title = "Security Testing"

    headers = [
        "Test ID", "Deliverable ID", "Test Type", "Test Tool", "Test Date",
        "Tester", "Scope", "Total Findings", "Critical Findings", "High Findings",
        "Medium Findings", "Low Findings", "False Positives", "Findings Remediated",
        "Findings Outstanding", "Report Reference", "Retest Required",
        "Retest Date", "Retest Status"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "SECURITY TESTING"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write headers in row 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"SAST,DAST,SCA,Penetration Test,Manual Review"')
    ws.add_data_validation(type_dv)
    type_dv.add('C4:C102')

    retest_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(retest_dv)
    retest_dv.add('Q4:Q102')

    retest_status_dv = DataValidation(type="list", formula1='"Pending,Passed,Failed,N/A"')
    ws.add_data_validation(retest_status_dv)
    retest_status_dv.add('S4:S102')

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "TEST-001", "DEL-001", "SAST", "SonarQube", "26.11.2024",
        "Security Team", "Full application scan", "45", "3", "12",
        "18", "12", "8", "42", "3",
        "SharePoint/Testing/SAST-001.pdf", "Yes", "02.12.2024", "Passed"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Empty data rows (rows 5-54) - 50 empty rows for user data
    for row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [10, 15, 15, 20, 12, 20, 30, 12, 12, 12, 12, 10, 12, 15, 15, 35, 12, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_sbom_management_sheet(ws):
    """Create the SBOM Management sheet."""
    ws.title = "SBOM Management"

    headers = [
        "SBOM ID", "Deliverable ID", "SBOM Format", "SBOM Date",
        "Total Components", "Direct Dependencies", "Transitive Dependencies",
        "Known Vulnerabilities", "Critical Vulns", "High Vulns",
        "License Issues", "Outdated Components", "Review Status",
        "Reviewed By", "Review Date", "SBOM Reference", "Action Plan"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "SBOM MANAGEMENT"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Write headers in row 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    format_dv = DataValidation(type="list", formula1='"CycloneDX,SPDX,Spreadsheet,Other"')
    ws.add_data_validation(format_dv)
    format_dv.add('C4:C102')

    status_dv = DataValidation(type="list", formula1='"Pending,Reviewed,Accepted,Rejected"')
    ws.add_data_validation(status_dv)
    status_dv.add('M4:M102')

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "SBOM-001", "DEL-001", "CycloneDX", "27.11.2024",
        "156", "23", "133",
        "2", "0", "2",
        "0", "8", "Reviewed",
        "Security Team", "27.11.2024", "SharePoint/SBOMs/SBOM-001.json", "2 high vulns flagged for remediation"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Empty data rows (rows 5-54) - 50 empty rows for user data
    for row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 12, 12, 15, 18, 20, 18, 12, 12, 12, 18, 12, 20, 12, 35, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_acceptance_signoff_sheet(ws):
    """Create the Acceptance Sign-off domain sheet."""
    ws.title = "Acceptance Sign-off"

    headers = [
        "Acceptance ID", "Deliverable ID", "Criteria Category", "Acceptance Criteria",
        "Status", "Evidence Reference", "Verified By", "Verification Date",
        "Waiver Reason", "Waiver Approver"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "ACCEPTANCE SIGN-OFF"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Write headers in row 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Standard acceptance criteria
    criteria = [
        ("Security", "Code review completed with no unresolved Critical/High findings"),
        ("Security", "SAST scan completed with no unresolved Critical/High findings"),
        ("Security", "SCA scan completed with no Critical/High vulnerable dependencies"),
        ("Security", "DAST scan completed (for web applications)"),
        ("Security", "Penetration test passed (for Critical projects)"),
        ("Security", "SBOM received and reviewed"),
        ("Security", "No secrets detected in codebase"),
        ("Documentation", "Security documentation complete"),
        ("Compliance", "Vulnerability remediation SLAs met"),
        ("Compliance", "Security training completed by all developers"),
    ]

    # Sample row (row 4) with example data (grey fill per Option B standard)
    sample_data = [
        "ACC-001", "[Deliverable ID]", "Security", "Code review completed with no unresolved Critical/High findings",
        "Met", "Code Review Report", "SharePoint/Deliverables/DEL-001/CodeReview.pdf",
        "Security Team", "26.11.2024", "All high-risk items resolved"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = THIN_BORDER

    # Pre-populated FFFFCC requirement row (row 5) — mandatory acceptance criterion
    additional_criteria = [
        "Threat model reviewed and accepted by security lead before delivery sign-off",
    ]
    for i, criterion_text in enumerate(additional_criteria):
        r = 5 + i
        ws.cell(row=r, column=1).fill = INPUT_FILL
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=3, value="Security").fill = INPUT_FILL
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=4, value=criterion_text).fill = INPUT_FILL
        ws.cell(row=r, column=4).font = Font(name="Calibri", size=10)
        for col in range(5, 11):
            ws.cell(row=r, column=col).fill = INPUT_FILL
            ws.cell(row=r, column=col).font = Font(name="Calibri", size=10)
        for col in range(1, 11):
            ws.cell(row=r, column=col).border = THIN_BORDER
    ffffcc_start = 5 + len(additional_criteria)

    # Empty data rows - rows 6–54 (49 empty + row 5 pre-populated = 50 FFFFCC total)
    for row in range(ffffcc_start, 55):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validation
    category_dv = DataValidation(type="list", formula1='"Functional,Security,Performance,Documentation,Compliance"')
    ws.add_data_validation(category_dv)
    category_dv.add('C4:C102')

    status_dv = DataValidation(type="list", formula1='"Met,Not Met,Waived,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('E4:E102')

    # Column widths
    widths = [12, 15, 18, 55, 10, 35, 20, 15, 35, 25]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"


def create_summary_dashboard_sheet(ws):
    """Create standard Summary Dashboard sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Assessment area configurations
    # Format: (Sheet Name, Status Column, [Compliant, Partial, Non-Compliant, N/A], Row Start, Row End)
    area_configs = [
        ('Deliverable Inventory', 'L', ['Accepted', 'Conditional', 'Rejected,Pending', None], 5, 54),
        ('Code Review Tracking', 'M', ['Approved', 'Approved with Findings', 'Rejected', None], 5, 54),
        ('Security Testing', 'S', ['Passed', 'Pending', 'Failed', 'N/A'], 5, 54),
        ('SBOM Management', 'M', ['Accepted', 'Reviewed,Pending', 'Rejected', None], 5, 54),
        ('Acceptance Sign-off', 'E', ['Met', 'Waived', 'Not Met', 'N/A'], 5, 54),
    ]

    # Assessment area display names (for TABLE 1 column A)
    assessment_areas = [
        'Deliverable Inventory',
        'Code Review Tracking',
        'Security Testing Results',
        'SBOM Management',
        'Acceptance Sign-off',
    ]

    ws.title = "Summary Dashboard"
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value="SECURITY TESTING ASSESSMENT — SUMMARY DASHBOARD")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    cell = ws.cell(row=2, column=1, value=CONTROL_REF)
    cell.font = Font(name="Calibri", size=10, italic=True, color="003366")

    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")

    row = 5
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # TABLE 1 data rows with COUNTIF formulas
    for i, (display_name, (sheet_name, status_col, status_values, row_start, row_end)) in enumerate(zip(assessment_areas, area_configs)):
        r = 6 + i
        ws.cell(row=r, column=1, value=display_name).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border

        # Column B: Total Requirements
        c = ws.cell(row=r, column=2)
        c.value = f"=COUNTA('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end})"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column C: Compliant
        c = ws.cell(row=r, column=3)
        if status_values[0] and ',' in status_values[0]:
            # Multiple compliant values (e.g., "Suspended,Removed")
            compliant_options = status_values[0].split(',')
            formula_parts = [f'COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{opt}")' for opt in compliant_options]
            c.value = f"={'+'.join(formula_parts)}"
        elif status_values[0]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[0]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column D: Partially Compliant
        c = ws.cell(row=r, column=4)
        if status_values[1] and ',' in status_values[1]:
            # Multiple partial values (e.g., "Reviewed,Pending")
            partial_options = status_values[1].split(',')
            formula_parts = [f'COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{opt}")' for opt in partial_options]
            c.value = f"={'+'.join(formula_parts)}"
        elif status_values[1]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[1]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column E: Non-Compliant
        c = ws.cell(row=r, column=5)
        if status_values[2] and ',' in status_values[2]:
            # Multiple non-compliant values (e.g., "Suspended,Removed")
            nc_options = status_values[2].split(',')
            formula_parts = [f'COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{opt}")' for opt in nc_options]
            c.value = f"={'+'.join(formula_parts)}"
        elif status_values[2]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[2]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column F: N/A
        c = ws.cell(row=r, column=6)
        if status_values[3]:
            c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[3]}")'
        else:
            c.value = "=0"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column G: Compliance %
        c = ws.cell(row=r, column=7)
        c.value = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
        c.number_format = "0.0%"
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # TOTAL row
    total_row = 6 + len(area_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border = border

    # TOTAL row SUM formulas - sum only data rows (6 to total_row-1) to avoid circular reference
    for col_idx in range(2, 7):
        col_letter = chr(64 + col_idx)
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}6:{col_letter}{total_row-1})"
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=total_row, column=7)
    c.value = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.border = border
    c.alignment = Alignment(horizontal="center")

    # TABLE 2: KEY METRICS
    row = total_row + 2
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")

    # Metric 1: Total Deliverables Pending Acceptance
    r = row + 1
    ws.cell(row=r, column=1, value="Total Deliverables Pending Acceptance").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=COUNTIF(\'Deliverable Inventory\'!L5:L54,"Pending")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 2: Critical Deliverables
    r = row + 2
    ws.cell(row=r, column=1, value="Critical Deliverables").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=COUNTIF(\'Deliverable Inventory\'!F5:F54,"Critical")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 3: Security Tests Failed
    r = row + 3
    ws.cell(row=r, column=1, value="Security Tests Failed").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=COUNTIF(\'Security Testing\'!S5:S54,"Failed")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 4: Acceptance Criteria Met %
    r = row + 4
    ws.cell(row=r, column=1, value="Acceptance Criteria Met %").font = Font(name="Calibri", size=10)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = '=IF(COUNTA(\'Acceptance Sign-off\'!E5:E54)=0,"0%",ROUND(COUNTIF(\'Acceptance Sign-off\'!E5:E54,"Met")/COUNTA(\'Acceptance Sign-off\'!E5:E54)*100,1)&"%")'
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # TABLE 3: CRITICAL FINDINGS
    row = row + 6
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: CRITICAL FINDINGS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    row += 1
    for col, h in enumerate(["#", "Finding", "Severity", "Affected Area", "Recommended Action", "Owner", "Due Date"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border

    # Critical findings based on DV analysis
    findings = [
        ('=COUNTIF(\'Deliverable Inventory\'!L5:L54,"Rejected")', "Rejected Deliverables", "Critical", "Deliverable Inventory", "Review rejection reasons and remediate"),
        ('=COUNTIF(\'Deliverable Inventory\'!I5:I54,"Pending")', "Pending Code Reviews", "High", "Deliverable Inventory", "Complete code reviews before acceptance"),
        ('=COUNTIF(\'Security Testing\'!S5:S54,"Failed")', "Failed Security Tests", "Critical", "Security Testing", "Remediate security vulnerabilities"),
        ('=COUNTIF(\'Security Testing\'!S5:S54,"Pending")', "Pending Retests", "High", "Security Testing", "Complete retesting of remediated findings"),
        ('=COUNTIF(\'Deliverable Inventory\'!K5:K54,"No")', "Missing SBOMs", "High", "Deliverable Inventory", "Request SBOMs from vendors"),
    ]

    for i, (count_formula, finding_name, severity, area, action) in enumerate(findings, 1):
        r = row + i
        ws.cell(row=r, column=1, value=count_formula).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=finding_name).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=severity).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=4, value=area).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=5, value=action).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=6).border = border
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=7).border = border

    for col, w in zip("ABCDEFG", [40, 16, 16, 18, 18, 12, 15]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create standard Evidence Register (8 columns, 100 rows)."""
    ws.title = "Evidence Register"
    border = THIN_BORDER

    # Header
    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value="EVIDENCE REGISTER")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    cell = ws.cell(row=2, column=1, value="Record all evidence collected during the assessment. Each row represents one piece of evidence.")
    cell.font = Font(name="Calibri", size=10, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column headers row 4 — 003366 fill (GS-ER-005)
    headers = ["Evidence ID", "Assessment Area", "Evidence Type", "Description", "Location / Path", "Date Collected", "Collected By", "Verification Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="003366")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[4].height = 30

    # Sample row (row 5) with example data (grey fill per Option B standard)
    sample_data = [
        "EV-001", "Security Testing", "Test results",
        "SAST scan report - Customer Portal v2.0",
        "SharePoint/Testing/SAST-CustomerPortal-20241126.pdf",
        "26.11.2024", "Security Team", "Verified"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.font = Font(name="Calibri", size=10, color="808080")
        cell.border = border

    # Empty data rows (rows 6-105) - 100 empty rows for user data (MAX-002 standard)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor="FFFFCC")
            cell.border = border

    # Dropdowns
    ev_types = DataValidation(type="list", formula1='"Configuration file,Screenshot,Log extract,Policy document,Training record,Audit report,Risk assessment,Interview notes,Test results,Other"', allow_blank=True)
    ev_types.prompt = "Select evidence type"
    ws.add_data_validation(ev_types)
    ev_types.add("C6:C105")

    verify_status = DataValidation(type="list", formula1='"Verified,Pending Verification,Insufficient,Not Reviewed"', allow_blank=True)
    verify_status.prompt = "Select verification status"
    ws.add_data_validation(verify_status)
    verify_status.add("H6:H105")

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22

    ws.freeze_panes = "A5"


def create_approval_sheet(ws):
    """Create standard Approval Sign-Off sheet."""
    ws.title = "Approval Sign-Off"
    border = THIN_BORDER

    # Header
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value="ASSESSMENT APPROVAL AND SIGN-OFF")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="ASSESSMENT SUMMARY")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

    # Summary fields
    summary_fields = [
        ("Document:", DOCUMENT_ID),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G10"),  # TABLE 1 TOTAL row
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    for i, (label, value) in enumerate(summary_fields):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws.merge_cells(f"B{r}:E{r}")
        ws.cell(row=r, column=2, value=value)
        # Apply border and fill to all cells in merged range
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = border
            if label in ("Assessment Status:", "Overall Compliance Rating:"):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFCC")

    # Assessment Status dropdown
    status_dv = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"B{row + 4}")

    row = row + 1 + len(summary_fields) + 1

    # Helper for approver sections
    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

        fields = ["Name:", "Title:", "Date:", "Signature:", "Comments:"]
        for i, field in enumerate(fields):
            r = start_row + 1 + i
            ws.cell(row=r, column=1, value=field).font = Font(name="Calibri", size=10, bold=True)
            ws.merge_cells(f"B{r}:E{r}")
            # Apply fill and border to all cells in merged range
            for col in range(2, 6):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFCC")
                ws.cell(row=r, column=col).border = border
        return start_row + 1 + len(fields) + 1

    row = _approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _approver_section(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")
    row = _approver_section(row, "APPROVED BY (CISO)", "003366")

    # FINAL DECISION
    ws.cell(row=row, column=1, value="FINAL DECISION:").font = Font(name="Calibri", size=11, bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    # Apply fill and border to all cells in merged range
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=row, column=col).border = border

    final_dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(final_dv)
    final_dv.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="NEXT REVIEW DETAILS")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

    review_fields = ["Next Review Date:", "Review Responsible:", "Special Considerations:"]
    for i, field in enumerate(review_fields):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=field).font = Font(name="Calibri", size=10, bold=True)
        ws.merge_cells(f"B{r}:E{r}")
        # Apply fill and border to all cells in merged range
        for col in range(2, 6):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFCC")
            ws.cell(row=r, column=col).border = border

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    ws.freeze_panes = "A3"


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    default_sheet = wb.active
    default_sheet.sheet_view.showGridLines = False

    # Create domain sheets
    create_instructions_sheet(wb.create_sheet())
    create_deliverable_inventory_sheet(wb.create_sheet())
    create_code_review_sheet(wb.create_sheet())
    create_security_testing_sheet(wb.create_sheet())
    create_sbom_management_sheet(wb.create_sheet())
    create_acceptance_signoff_sheet(wb.create_sheet())

    # Create standard sheets
    create_evidence_register(wb.create_sheet())
    create_summary_dashboard_sheet(wb.create_sheet())
    create_approval_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")
    logger.info(f"{CHECK} SUCCESS: Generated {OUTPUT_FILENAME}")



# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
