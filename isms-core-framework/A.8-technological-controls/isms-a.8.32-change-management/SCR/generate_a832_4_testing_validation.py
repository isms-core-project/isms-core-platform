#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.32.4 - Testing & Validation Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Assessment Domain 4 of 4: Testing, Validation & Acceptance Procedures

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific testing frameworks and validation requirements.

Key customization areas:
1. Testing framework and methodologies (match your SDLC)
2. Test types and coverage requirements (adapt to your risk profile)
3. Acceptance criteria definitions (align with your quality gates)
4. Rollback procedures (customize to your infrastructure)
5. Production validation methods (specific to your systems)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
testing procedures, validation processes, acceptance criteria, and rollback
capabilities against ISO 27001:2022 Control A.8.32 requirements.

**Purpose:**
Enables systematic assessment of testing frameworks, validation procedures,
acceptance criteria, rollback capabilities, and production validation to ensure
changes are properly tested before deployment and can be reversed if needed.

**Assessment Scope:**
- Testing framework overview and methodologies
- Unit and integration testing procedures
- User acceptance testing (UAT) and business validation
- Security and regression testing integration
- Acceptance criteria definition and sign-off procedures
- Rollback procedure testing and validation
- Production validation and verification procedures
- Test coverage and effectiveness measurement
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and testing standards
2. Testing_Framework_Overview - Framework and methodology documentation
3. Unit_Integration_Testing - Developer testing procedures
4. UAT_Business_Validation - User acceptance and business validation
5. Security_Regression_Testing - Security and regression test integration
6. Acceptance_Criteria - Quality gates and acceptance definitions
7. Summary_Dashboard - Compliance metrics and analytics
8. Evidence_Register - Audit evidence tracking
9. Approval_Sign_Off - Stakeholder approval workflow

**Key Features:**
- Technology-agnostic assessment (any testing tools/frameworks)
- Test type coverage matrix (unit, integration, UAT, security, regression)
- Acceptance criteria templates and examples
- Rollback procedure validation and testing
- Automated compliance calculations
- Evidence linkage for audit traceability

**Integration:**
with Control 8.29 (Security Testing in Development and Acceptance).

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a832_4_testing_validation.py

Output:
    File: ISMS_A_8_32_4_Testing_Validation_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Document your organisation's testing framework
    2. Define test types and coverage requirements
    3. Establish acceptance criteria for different change types
    4. Document rollback procedures and testing
    5. Define production validation procedures
    6. Review Summary_Dashboard for compliance metrics
    7. Collect and link test evidence

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32, A.8.29
Assessment Domain:    4 of 4 (Testing, Validation & Acceptance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.4: Testing & Validation Implementation Guide
    - ISMS-IMP-A.8.32.1: Change Process Assessment (Domain 1)
    - ISMS-IMP-A.8.32.2: Change Types & Categories Assessment (Domain 2)
    - ISMS-IMP-A.8.32.3: Environment Separation Assessment (Domain 3)
    - ISMS-POL-A.8.29: Security Testing in Development and Acceptance

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.32.4 specification
    - Supports comprehensive testing and validation evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Testing is Risk Mitigation:**
Comprehensive testing reduces the likelihood of change failures, production
incidents, and security vulnerabilities. Testing is not bureaucracy - it's
risk management.

**Test Coverage vs. 100% Testing:**
Not all changes require all test types. Risk-based testing is appropriate.
Standard changes may need less testing than high-risk normal changes.

**Rollback is Not Optional:**
Every production change must have a documented and tested rollback procedure.
"Hope for the best" is not a rollback plan.

**Production Validation Matters:**
Post-deployment validation in production is essential. Did the change actually
work as expected? Are there performance impacts? Are users affected?

**Audit Considerations:**
Auditors will verify testing procedures exist, are followed, and generate
evidence. They'll look for test plans, test results, and acceptance sign-offs.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.32.4"
WORKBOOK_NAME = "Testing & Validation Assessment"
CONTROL_ID = "A.8.32"
CONTROL_NAME = "Change Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠  Warning sign
GEAR = '\u2699'       # ⚙  Gear
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP-A.8.32.4 spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.32.4 specification (11 sheets)
    sheets = [
        "Instructions & Legend",
        "Testing Framework",
        "Test Coverage",
        "Acceptance Criteria",
        "Rollback Procedures",
        "Production Validation",
        "Security Testing",
        "Testing Documentation",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create ALL data validation objects for dropdowns.
    CRITICAL: Must include EVERY dropdown type used in the spec.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,❌ No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,❌ No,N/A"',
            allow_blank=False
        ),
        'yes_partial_no': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,⚠ Partial,❌ No"',
            allow_blank=False
        ),
        'yes_partial_no_na': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,⚠ Partial,❌ No,N/A"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Implemented,⚠ Partial,❌ Not Implemented, Planned,N/A"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Compliant,⚠ Partial,❌ Non-Compliant, Pending"',
            allow_blank=False
        ),
        'automation_level': DataValidation(
            type="list",
            formula1=f'"{CHECK} Automated,⚠ Semi-Automated,❌ Manual"',
            allow_blank=False
        ),
        'integration_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Integrated,⚠ Partial,❌ Standalone"',
            allow_blank=False
        ),
        'license_status': DataValidation(
            type="list",
            formula1='"Active,Expiring Soon,Expired"',
            allow_blank=False
        ),
        'testing_required': DataValidation(
            type="list",
            formula1=f'"{CHECK} Mandatory,⚠ Recommended,❌ Optional"',
            allow_blank=False
        ),
        'test_result': DataValidation(
            type="list",
            formula1=f'"{CHECK} Pass,❌ Fail,⚠ Partial, Pending"',
            allow_blank=False
        ),
        'test_frequency': DataValidation(
            type="list",
            formula1='"Every Deploy,Daily,Weekly,Monthly,Per Release,On-Demand"',
            allow_blank=False
        ),
        'critical_level': DataValidation(
            type="list",
            formula1=f'"{CHECK} Critical,⚠ High, Medium,N/A"',
            allow_blank=False
        ),
        'rollback_method': DataValidation(
            type="list",
            formula1=f'"{CHECK} Automated,⚠ Semi-Automated,❌ Manual,N/A"',
            allow_blank=False
        ),
        'test_regularly': DataValidation(
            type="list",
            formula1=f'"{CHECK} Regularly,⚠ Occasionally,❌ Never"',
            allow_blank=False
        ),
        'signoff_method': DataValidation(
            type="list",
            formula1='"Electronic,Written,Verbal,Automated"',
            allow_blank=False
        ),
        'detection_method': DataValidation(
            type="list",
            formula1='"Automated,Manual,User Report"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Change Request,Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other,N/A"',
            allow_blank=True
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⚠ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Final,⚠ Requires Remediation, Draft,❌ Re-assessment Required"',
            allow_blank=False
        ),
        'review_recommendation': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approve,⚠ Approve with Conditions,❌ Reject, Require Rework"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approved,⚠ Approved with Conditions,❌ Rejected"',
            allow_blank=False
        ),
    }
    
    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Document YOUR testing framework and methodologies.', '2. Assess test coverage across different test types (unit, integration, UAT, security, etc.).', '3. Define YOUR acceptance criteria for change validation.', '4. Document YOUR rollback procedures and testing.', '5. Assess YOUR production validation processes.', '6. Review the Summary Dashboard for compliance metrics.', '7. Maintain the Evidence Register for audit traceability.', '8. Obtain final approval via Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_testing_framework_assessment(ws, styles):
    """Create Testing_Framework_Assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "TESTING FRAMEWORK ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = "Document testing strategy, methodologies, and governance"
    apply_style(ws["A2"], styles["subheader"])

    # Testing Strategy & Governance
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TESTING STRATEGY & GOVERNANCE"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Aspect", "Current State", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    strategy_aspects = [
        "Testing strategy documented?",
        "Testing policy established?",
        "Test process integrated into SDLC?",
        "Testing roles & responsibilities defined?",
        "Test environment management defined?",
        "Test data management procedures?",
        "Defect management process defined?",
        "Test metrics & KPIs defined?",
        "Test automation strategy?",
        "Continuous testing implemented?",
        "Test tool standardization?",
        "Testing training program?",
        "Test governance board/committee?",
        "Third-party/vendor testing requirements?",
        "Compliance testing procedures (regulatory)?",
    ]

    row += 1
    for aspect in strategy_aspects:
        ws[f"A{row}"] = aspect

        # Current State dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        # Compliance dropdown
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        # Evidence
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=4))

        # Notes
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        row += 1

    # Testing Tools Inventory
    row += 1
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "TESTING TOOLS INVENTORY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    tool_headers = ["Tool Category", "Tool Name", "Purpose", "Automation Level", "Integration", "License Status", "Owner", "Compliance", "Evidence"]
    for col_idx, header in enumerate(tool_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    tool_categories = [
        "Unit Testing",
        "Integration Testing",
        "Functional Testing",
        "UAT/Acceptance Testing",
        "Security Testing (SAST)",
        "Security Testing (DAST)",
        "Dependency Scanning (SCA)",
        "Performance Testing",
        "Load Testing",
        "Regression Testing",
        "API Testing",
        "UI/E2E Testing",
        "Test Management",
        "Defect Tracking",
        "Test Data Management",
    ]

    row += 1
    for category in tool_categories:
        ws[f"A{row}"] = category
        ws[f"A{row}"].font = Font(bold=True)
        
        # Tool Name
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        # Purpose
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Automation Level
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['automation_level'].add(ws[f"D{row}"])

        # Integration
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['integration_status'].add(ws[f"E{row}"])

        # License Status
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['license_status'].add(ws[f"F{row}"])

        # Owner
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        # Compliance
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"H{row}"])

        # Evidence
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"I{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=9))

        row += 1

    # Testing Governance Metrics
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TESTING GOVERNANCE METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Target Value", "Current Value", "Status", "Notes"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        ("Overall test automation rate", ">70%"),
        ("Test coverage (code coverage)", ">80%"),
        ("Defect detection rate (pre-prod)", ">95%"),
        ("Test execution time (CI/CD)", "<30 min"),
        ("Test pass rate (first run)", ">90%"),
        ("Critical defects escaped to production", "<2 per quarter"),
        ("Mean time to detect defects (MTTD)", "<24 hours"),
        ("Test environment availability", ">95%"),
        ("Rollback success rate", "100%"),
    ]

    row += 1
    for metric_name, target in metrics:
        ws[f"A{row}"] = metric_name
        ws[f"B{row}"] = target
        
        # Current Value
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Status - editable (manual assessment; targets are text-based)
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        # Notes
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 25

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: TEST_TYPES_COVERAGE SHEET
# ============================================================================

def create_test_types_coverage(ws, styles):
    """Create Test_Types_Coverage sheet with comprehensive test type assessment."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "TEST TYPES & COVERAGE ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Assess implementation and coverage of different test types"
    apply_style(ws["A2"], styles["subheader"])

    # Common headers for test sections
    test_headers = ["Aspect", "Implemented?", "Coverage/Details", "Automation Level", "Tool/Framework", "Compliance", "Evidence"]

    # Unit Testing Assessment
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "UNIT TESTING ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    unit_test_aspects = [
        "Unit test framework established?",
        "Code coverage measurement?",
        "Minimum coverage threshold enforced?",
        "Unit tests run in CI/CD pipeline?",
        "Unit test failures block deployment?",
        "Mock/stub frameworks used?",
        "Test data generation automated?",
        "Code review includes test review?",
        "TDD/BDD practices followed?",
        "Unit test documentation maintained?",
        "Unit test metrics tracked?",
    ]

    row += 1
    for aspect in unit_test_aspects:
        ws[f"A{row}"] = aspect

        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['automation_level'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # Integration Testing Assessment
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "INTEGRATION TESTING ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    integration_aspects = [
        "Integration test strategy defined?",
        "API integration testing?",
        "Database integration testing?",
        "Third-party service integration testing?",
        "Microservices integration testing?",
        "Message queue/event testing?",
        "End-to-end integration scenarios?",
        "Integration test environment availability?",
        "Test data management for integration?",
        "Integration tests in CI/CD?",
        "Integration test failures block deployment?",
    ]

    row += 1
    for aspect in integration_aspects:
        ws[f"A{row}"] = aspect
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['automation_level'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # System/Functional Testing
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "SYSTEM/FUNCTIONAL TESTING ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    functional_aspects = [
        "Functional test cases documented?",
        "Business requirements coverage?",
        "UI/UX testing performed?",
        "Cross-browser testing?",
        "Mobile/responsive testing?",
        "Accessibility testing (WCAG)?",
        "Localization/i18n testing?",
        "Regression test suite maintained?",
        "Regression tests automated?",
        "Smoke/sanity test suite?",
        "Exploratory testing performed?",
    ]

    row += 1
    for aspect in functional_aspects:
        ws[f"A{row}"] = aspect
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['automation_level'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # UAT Assessment
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "USER ACCEPTANCE TESTING (UAT) ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    uat_aspects = [
        "UAT process defined and documented?",
        "UAT environment available?",
        "UAT test scenarios based on user stories?",
        "Business users involved in UAT?",
        "UAT sign-off process defined?",
        "UAT sign-off required before production?",
        "UAT test data management?",
        "UAT defects tracked separately?",
        "UAT feedback incorporated?",
        "UAT documentation maintained?",
    ]

    row += 1
    for aspect in uat_aspects:
        ws[f"A{row}"] = aspect
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if "sign-off required" in aspect.lower():
            validations['testing_required'].add(ws[f"B{row}"])
        else:
            validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['automation_level'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # Security Testing (Control 8.29)
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "SECURITY TESTING ASSESSMENT (CONTROL 8.29 - CRITICAL)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    security_aspects = [
        "Security testing strategy defined?",
        "Static Application Security Testing (SAST)?",
        "Dynamic Application Security Testing (DAST)?",
        "Software Composition Analysis (SCA)?",
        "Dependency vulnerability scanning?",
        "Container/image scanning?",
        "Infrastructure as Code (IaC) scanning?",
        "Secret detection/scanning?",
        "API security testing?",
        "Authentication/authorisation testing?",
        "SQL injection testing?",
        "XSS vulnerability testing?",
        "CSRF vulnerability testing?",
        "Security headers validation?",
        "Encryption/TLS testing?",
        "OWASP Top 10 testing coverage?",
        "Security test results block deployment?",
        "Security findings remediation tracking?",
        "Penetration testing performed?",
        "Security test reporting to security team?",
    ]

    row += 1
    for aspect in security_aspects:
        ws[f"A{row}"] = aspect
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['automation_level'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # Performance Testing
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "PERFORMANCE TESTING ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    performance_aspects = [
        "Performance testing strategy defined?",
        "Performance requirements defined?",
        "Load testing performed?",
        "Stress testing performed?",
        "Spike testing performed?",
        "Endurance/soak testing?",
        "Scalability testing?",
        "Response time monitoring?",
        "Throughput/TPS measurement?",
        "Resource utilization monitoring?",
        "Database performance testing?",
        "API performance testing?",
        "CDN/caching performance testing?",
        "Performance baselines established?",
        "Performance regression testing?",
        "Performance test results analysed?",
        "Performance issues remediated?",
    ]

    row += 1
    for aspect in performance_aspects:
        ws[f"A{row}"] = aspect
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['automation_level'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: ACCEPTANCE_CRITERIA_MANAGEMENT SHEET
# ============================================================================

def create_acceptance_criteria_management(ws, styles):
    """Create Acceptance_Criteria_Management sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "ACCEPTANCE CRITERIA MANAGEMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Define and manage acceptance criteria for change validation"
    apply_style(ws["A2"], styles["subheader"])

    # Acceptance Criteria Framework
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "ACCEPTANCE CRITERIA FRAMEWORK"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Aspect", "Implemented?", "Details", "Ownership", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    framework_aspects = [
        "Acceptance criteria definition process?",
        "Criteria defined at requirements phase?",
        "Measurable/testable criteria required?",
        "Functional acceptance criteria template?",
        "Non-functional criteria template?",
        "Security acceptance criteria?",
        "Performance acceptance criteria?",
        "Accessibility criteria (WCAG)?",
        "Compliance/regulatory criteria?",
        "Data protection/privacy criteria?",
        "Criteria review process?",
        "Criteria traceability to requirements?",
        "Acceptance criteria repository/tool?",
    ]

    row += 1
    for aspect in framework_aspects:
        ws[f"A{row}"] = aspect
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"E{row}"])

        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=6))

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        row += 1

    # Additional ISO 27002:2022 acceptance criteria (F2F2F2 reference rows)
    extra_aspects = [
        "BCM/DR plans reviewed and updated — for changes to critical or high-availability systems, continuity plans verified as current",
    ]
    for aspect in extra_aspects:
        ws.cell(row=row, column=1).value = aspect
        ws.cell(row=row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, color="003366")
        row += 1

    # Change Type Acceptance Criteria
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CHANGE TYPE ACCEPTANCE CRITERIA"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    change_headers = ["Change Type", "Mandatory Criteria", "Optional Criteria", "Sign-Off Required", "Compliance", "Evidence"]
    for col_idx, header in enumerate(change_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    change_types = [
        "Standard Change",
        "Normal Change - Low Risk",
        "Normal Change - Medium Risk",
        "Normal Change - High Risk",
        "Emergency Change",
        "Infrastructure Change",
        "Application Change",
        "Database Change",
        "Security Patch",
        "Configuration Change",
        "New Feature/Enhancement",
        "Bug Fix",
    ]

    row += 1
    for change_type in change_types:
        ws[f"A{row}"] = change_type
        ws[f"A{row}"].font = Font(bold=True)
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"E{row}"])

        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=6))

        row += 1

    # Acceptance Testing Stages
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "ACCEPTANCE TESTING STAGES"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    stage_headers = ["Stage", "Purpose", "Entry Criteria", "Exit Criteria", "Owner", "Duration", "Compliance", "Evidence"]
    for col_idx, header in enumerate(stage_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    testing_stages = [
        "Developer Testing",
        "Code Review",
        "Build Verification Testing (BVT)",
        "Integration Testing",
        "System Testing",
        "Security Testing",
        "Performance Testing",
        "UAT (Business)",
        "UAT (Technical)",
        "Regression Testing",
        "Pre-Production Validation",
        "Production Validation",
    ]

    row += 1
    for stage in testing_stages:
        ws[f"A{row}"] = stage
        ws[f"A{row}"].font = Font(bold=True)
        
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"G{row}"])

        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # Acceptance Criteria Tracking
    row += 1
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "ACCEPTANCE CRITERIA TRACKING (20 ROWS)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    tracking_headers = ["Criteria ID", "Change/Release", "Criteria Description", "Type", "Expected Result", "Actual Result", "Status", "Validated By", "Date", "Evidence"]
    for col_idx, header in enumerate(tracking_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Create 20 tracking rows
    row += 1
    for i in range(1, 21):
        ws[f"A{row}"] = f"AC-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        for col in ["B", "C", "D", "E", "F", "H", "J"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Status dropdown
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['test_result'].add(ws[f"G{row}"])

        # Date
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"I{row}"].border = styles["border"]
        ws[f"I{row}"].number_format = 'DD.MM.YYYY'

        # Evidence dropdown
        validations['evidence_type'].add(ws.cell(row=row, column=10))

        # Alternating row colors
        if i % 2 == 0:
            for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
                if ws[f"{col}{row}"].fill.start_color.rgb != "FFFFCC":
                    ws[f"{col}{row}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 25

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# Let me continue with the remaining sections in the next part...
# ============================================================================
# SECTION 7: ROLLBACK_PROCEDURES SHEET
# ============================================================================

def create_rollback_procedures(ws, styles):
    """Create Rollback_Procedures assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ROLLBACK PROCEDURES ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = "Assess rollback capability and procedures for safe change reversals"
    apply_style(ws["A2"], styles["subheader"])

    # Rollback Strategy & Governance
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ROLLBACK STRATEGY & GOVERNANCE"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Aspect", "Implemented?", "Details", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    strategy_aspects = [
        "Rollback strategy documented?",
        "Rollback procedures mandatory for all changes?",
        "Rollback decision criteria defined?",
        "Rollback authorisation process?",
        "Rollback testing required pre-deployment?",
        "Rollback time objectives (RTO) defined?",
        "Rollback success metrics tracked?",
        "Failed rollback escalation procedure?",
        "Post-rollback validation required?",
        "Rollback communication procedures?",
        "Rollback documentation maintained?",
        "Rollback training provided?",
    ]

    row += 1
    for aspect in strategy_aspects:
        ws[f"A{row}"] = aspect
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=5))

        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        row += 1

    # Rollback Capability by Change Type
    row += 1
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "ROLLBACK CAPABILITY BY CHANGE TYPE"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    capability_headers = ["Change Type", "Rollback Method", "Automated?", "Tested?", "RTO Target", "Last Test Date", "Success Rate", "Compliance", "Evidence"]
    for col_idx, header in enumerate(capability_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    change_types = [
        "Application Deployment",
        "Database Schema Change",
        "Database Data Change",
        "Configuration Change",
        "Infrastructure Change",
        "Network Change",
        "Security Patch",
        "Cloud Resource Change",
        "Container/K8s Deployment",
        "Serverless Function",
        "API Gateway Change",
        "Load Balancer Change",
        "DNS Change",
        "Certificate Change",
        "Firewall Rule Change",
        "IAM Policy Change",
        "Monitoring/Alerting Change",
        "CI/CD Pipeline Change",
    ]

    row += 1
    for change_type in change_types:
        ws[f"A{row}"] = change_type
        ws[f"A{row}"].font = Font(bold=True)
        
        # Rollback Method
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['rollback_method'].add(ws[f"B{row}"])

        # Automated?
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"C{row}"])

        # Tested?
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['test_regularly'].add(ws[f"D{row}"])

        # RTO Target
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        # Last Test Date
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        ws[f"F{row}"].number_format = 'DD.MM.YYYY'

        # Success Rate
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        # Compliance
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"H{row}"])

        # Evidence
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"I{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=9))

        row += 1

    # Rollback Testing History
    row += 1
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "ROLLBACK TESTING HISTORY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    test_headers = ["Test ID", "Change/Release", "Change Type", "Rollback Method", "Test Date", "Test Result", "Duration", "Issues Identified", "Remediation", "Evidence"]
    for col_idx, header in enumerate(test_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Sample row with example data
    row += 1
    sample_data = {
        1: "RB-001",
        2: "CHG-2025-001",
        3: "Application Deployment",
        4: "Automated Rollback",
        5: "15.01.2026",
        6: "Success",
        7: "12 minutes",
        8: "None",
        9: "N/A",
        10: "EV-RB-001",
    }
    for col_idx, value in sample_data.items():
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = styles["border"]
        if col_idx == 5:  # Date column
            cell.number_format = 'DD.MM.YYYY'

    # Add dropdowns to sample row
    validations['rollback_method'].add(ws[f"D{row}"])
    validations['test_result'].add(ws[f"F{row}"])
    validations['evidence_type'].add(ws.cell(row=row, column=10))

    # Empty data rows (50 rows for user data)
    row += 1
    for i in range(50):
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]
            cell.value = None  # Empty - users document their own rollback tests
            if col_idx == 5:  # Date column
                cell.number_format = 'DD.MM.YYYY'

        # Add dropdowns to empty rows
        validations['rollback_method'].add(ws[f"D{row}"])
        validations['test_result'].add(ws[f"F{row}"])
        validations['evidence_type'].add(ws.cell(row=row, column=10))

        row += 1

    # Rollback Metrics & Analysis
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ROLLBACK METRICS & ANALYSIS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Target", "Current", "Status", "Notes"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        ("Rollbacks as % of total changes", "<5%"),
        ("Successful rollback rate", ">98%"),
        ("Average rollback time", "<30 min"),
        ("Rollbacks requiring escalation", "<10%"),
        ("Rollback testing coverage", "100%"),
        ("Automated rollback rate", ">80%"),
        ("Rollback-related incidents", "0 critical"),
        ("Data loss incidents during rollback", "0"),
    ]

    row += 1
    for metric_name, target in metrics:
        ws[f"A{row}"] = metric_name
        ws[f"B{row}"] = target
        
        # Current
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Status - editable (manual assessment; targets are text-based)
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        # Notes
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 25
    ws.column_dimensions["J"].width = 25

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: PRODUCTION_VALIDATION SHEET
# ============================================================================

def create_production_validation(ws, styles):
    """Create Production_Validation assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "PRODUCTION VALIDATION ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = "Assess post-deployment validation and monitoring in production"
    apply_style(ws["A2"], styles["subheader"])

    # Production Validation Strategy
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "PRODUCTION VALIDATION STRATEGY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Aspect", "Implemented?", "Details", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    strategy_aspects = [
        "Production validation process documented?",
        "Validation required for all deployments?",
        "Validation checklist template?",
        "Automated validation checks?",
        "Smoke test suite for production?",
        "Critical path validation?",
        "User acceptance in production (UAT-P)?",
        "Production monitoring integration?",
        "Validation timeframe defined?",
        "Failed validation rollback trigger?",
        "Validation sign-off required?",
        "Post-validation report required?",
    ]

    row += 1
    for aspect in strategy_aspects:
        ws[f"A{row}"] = aspect

        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=5))

        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        row += 1

    # Validation Checks by Category (35+ checks)
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "VALIDATION CHECKS BY CATEGORY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    check_headers = ["Check Category", "Check Description", "Automated?", "Critical?", "Owner", "Frequency", "Compliance", "Evidence"]
    for col_idx, header in enumerate(check_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Deployment Verification
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "** DEPLOYMENT VERIFICATION **"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
    
    deployment_checks = [
        "Deployment successful (all components)",
        "Service/application started successfully",
        "Database migrations completed",
        "Configuration applied correctly",
        "Dependencies/libraries loaded",
    ]

    row += 1
    for check in deployment_checks:
        ws[f"A{row}"] = "Deployment"
        ws[f"B{row}"] = check
        
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        validations['yes_partial_no'].add(ws[f"C{row}"])
        validations['critical_level'].add(ws[f"D{row}"])
        validations['test_frequency'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"G{row}"])

        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # Functional Validation
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "** FUNCTIONAL VALIDATION **"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
    
    functional_checks = [
        "Critical business flows operational",
        "User authentication working",
        "User authorisation working",
        "API endpoints responding",
        "Database queries executing",
        "External integrations working",
        "Payment processing (if applicable)",
    ]

    row += 1
    for check in functional_checks:
        ws[f"A{row}"] = "Functional"
        ws[f"B{row}"] = check
        
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        validations['yes_partial_no'].add(ws[f"C{row}"])
        validations['critical_level'].add(ws[f"D{row}"])
        validations['test_frequency'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"G{row}"])

        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # Performance Validation
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "** PERFORMANCE VALIDATION **"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
    
    performance_checks = [
        "Response time within SLA",
        "Throughput within expected range",
        "Resource utilization normal",
        "Database performance acceptable",
        "Memory usage stable",
    ]

    row += 1
    for check in performance_checks:
        ws[f"A{row}"] = "Performance"
        ws[f"B{row}"] = check
        
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        validations['yes_partial_no'].add(ws[f"C{row}"])
        validations['critical_level'].add(ws[f"D{row}"])
        validations['test_frequency'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"G{row}"])

        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # Security Validation
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "** SECURITY VALIDATION **"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
    
    security_checks = [
        "SSL/TLS certificates valid",
        "Security headers present",
        "No security misconfigurations",
        "Access controls working",
        "Audit logging functional",
    ]

    row += 1
    for check in security_checks:
        ws[f"A{row}"] = "Security"
        ws[f"B{row}"] = check
        
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        validations['yes_partial_no'].add(ws[f"C{row}"])
        validations['critical_level'].add(ws[f"D{row}"])
        validations['test_frequency'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"G{row}"])

        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # Monitoring & Alerting
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "** MONITORING & ALERTING **"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
    
    monitoring_checks = [
        "Application monitoring active",
        "Error tracking functional",
        "Performance monitoring active",
        "Alerts configured correctly",
        "Log aggregation working",
    ]

    row += 1
    for check in monitoring_checks:
        ws[f"A{row}"] = "Monitoring"
        ws[f"B{row}"] = check
        
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        validations['yes_partial_no'].add(ws[f"C{row}"])
        validations['critical_level'].add(ws[f"D{row}"])
        validations['test_frequency'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"G{row}"])

        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # User Impact Validation
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "** USER IMPACT VALIDATION **"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
    
    user_impact_checks = [
        "No user-reported issues",
        "Error rate within normal range",
        "No unexpected user behaviour",
        "Customer satisfaction maintained",
    ]

    row += 1
    for check in user_impact_checks:
        ws[f"A{row}"] = "User Impact"
        ws[f"B{row}"] = check
        
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        validations['yes_partial_no'].add(ws[f"C{row}"])
        validations['critical_level'].add(ws[f"D{row}"])
        validations['test_frequency'].add(ws[f"F{row}"])

        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"G{row}"])

        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"H{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 7.5: SECURITY TESTING SHEET
# ============================================================================

def create_security_testing(ws, styles):
    """Create Security_Testing sheet assessing security testing practices."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "SECURITY TESTING ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Assess security testing practices for change validation"
    apply_style(ws["A2"], styles["subheader"])

    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "SECURITY TESTING TYPES"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    headers = ["Test Type", "Description", "When Required", "Performed By", "Frequency", "Status", "Evidence"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    tests = [
        ("SAST (Static Analysis)", "Code security scanning", "All code changes", "", "", "", ""),
        ("DAST (Dynamic Analysis)", "Runtime vulnerability testing", "Web/API changes", "", "", "", ""),
        ("Dependency Scanning", "Third-party vulnerability check", "Library changes", "", "", "", ""),
        ("Penetration Testing", "Simulated attack testing", "Major releases", "", "", "", ""),
        ("Secret Scanning", "Credential leak detection", "All commits", "", "", "", ""),
        ("Container Scanning", "Image vulnerability analysis", "Container changes", "", "", "", ""),
        ("Infrastructure Scanning", "IaC security validation", "Infra changes", "", "", "", ""),
    ]

    for test_data in tests:
        for col_idx, value in enumerate(test_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 3:
                apply_style(cell, styles["input_cell"])
            # Add compliance DV to Status column (F)
            if col_idx == 6:
                validations['compliance_status'].add(cell)
            # Add evidence_type DV to Evidence column (G)
            if col_idx == 7:
                validations['evidence_type'].add(cell)
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: TESTING DOCUMENTATION SHEET
# ============================================================================

def create_testing_documentation(ws, styles):
    """Create Testing_Documentation sheet assessing test documentation quality."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "TESTING DOCUMENTATION ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Assess test documentation completeness and quality"
    apply_style(ws["A2"], styles["subheader"])

    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "DOCUMENTATION REQUIREMENTS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    headers = ["Document Type", "Description", "Template Exists", "Consistently Used", "Review Process", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    docs = [
        ("Test Plan", "Overall testing strategy and scope", "", "", "", ""),
        ("Test Cases", "Detailed test scenarios and steps", "", "", "", ""),
        ("Test Results", "Execution outcomes and defects", "", "", "", ""),
        ("Traceability Matrix", "Requirements to tests mapping", "", "", "", ""),
        ("UAT Sign-Off", "Business acceptance documentation", "", "", "", ""),
        ("Performance Reports", "Load/stress test results", "", "", "", ""),
        ("Security Reports", "Vulnerability scan results", "", "", "", ""),
    ]

    for doc_data in docs:
        for col_idx, value in enumerate(doc_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles["input_cell"])
            # Add compliance DV to Status column (F)
            if col_idx == 6:
                validations['compliance_status'].add(cell)
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: SUMMARY_DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with standard compliance table and metrics."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "TESTING & VALIDATION ASSESSMENT — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.32: Change Management"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # --- TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW ---
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    assessment_areas = [
        "Testing Framework",
        "Test Coverage",
        "Acceptance Criteria",
        "Rollback Strategy",
        "Rollback Capability",
        "Production Validation",
        "Security Testing",
        "Testing Documentation",
    ]

    # Area configurations: (sheet_name, status_col, [compliant, partial, non_compliant], row_start, row_end)
    # None = manual entry area
    area_configs = [
        ('Testing Framework', 'C', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 55),
        ('Test Coverage', 'F', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 100),  # Multiple test type sections (Unit, Integration, etc.)
        ('Acceptance Criteria', 'E', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 71),  # Multiple acceptance criteria sections (extended for BCM/DR row)
        ('Rollback Procedures', 'D', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 103),  # Section 1: Strategy & Governance (extended to cover all data)
        ('Rollback Procedures', 'H', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 21, 103),  # Section 2: Capability by Change Type (extended to cover all data)
        ('Production Validation', 'D', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 62),  # Production validation assessment sections
        ('Security Testing', 'F', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 12),  # Security testing types (7 tests)
        ('Testing Documentation', 'F', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 12),  # Documentation requirements (7 docs)
    ]

    for i, area_name in enumerate(assessment_areas):
        r = 6 + i
        ws[f"A{r}"] = area_name
        ws[f"A{r}"].font = Font(name="Calibri", size=10)
        ws[f"A{r}"].border = border

        if area_configs[i] is None:
            # MANUAL ENTRY - use placeholder cells
            for col in "BCDEF":
                cell = ws[f"{col}{r}"]
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Manual areas get placeholder in column G
            ws[f"G{r}"] = "[enter %]"
            ws[f"G{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        else:
            # FORMULA-BASED - generate compliance formulas
            sheet_name, status_col, status_values, row_start, row_end = area_configs[i]
            compliant_val, partial_val, non_compliant_val = status_values

            # B: Total items (count non-empty rows in column A)
            ws[f"B{r}"] = f"=COUNTA('{sheet_name}'!A{row_start}:A{row_end})"
            ws[f"B{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"B{r}"].border = border
            ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # C: Compliant
            ws[f"C{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{compliant_val}\")"
            ws[f"C{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"C{r}"].border = border
            ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # D: Partially Compliant
            ws[f"D{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{partial_val}\")"
            ws[f"D{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"D{r}"].border = border
            ws[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # E: Non-Compliant (including "Pending" if present in DV)
            ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")+COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"Pending\")"
            ws[f"E{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"E{r}"].border = border
            ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # F: N/A (placeholder - no N/A status in these DVs)
            ws[f"F{r}"] = 0
            ws[f"F{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"F{r}"].border = border
            ws[f"F{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # G: Compliance % = Compliant / (Total - N/A) * 100
            ws[f"G{r}"] = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
            ws[f"G{r}"].number_format = "0.0%"

        # Apply border and alignment to G column (all rows)
        ws[f"G{r}"].border = border
        ws[f"G{r}"].alignment = Alignment(horizontal="center", vertical="center")

    # TOTAL row with SUM formulas
    total_row = 6 + len(assessment_areas)
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{total_row}"].border = border
    for col_letter in "BCDEF":
        cell = ws[f"{col_letter}{total_row}"]
        cell.value = f"=SUM({col_letter}6:{col_letter}{total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"G{total_row}"] = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    ws[f"G{total_row}"].number_format = "0.0%"
    ws[f"G{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"G{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"G{total_row}"].border = border
    ws[f"G{total_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # --- TABLE 2: KEY METRICS ---
    met_row = total_row + 2
    ws.merge_cells(f"A{met_row}:G{met_row}")
    ws[f"A{met_row}"] = "TABLE 2: KEY METRICS"
    ws[f"A{met_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{met_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{met_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=met_row, column=col).border = border

    # TABLE 2 column headers (D9D9D9 grey)
    t2_hdr_row = met_row + 1
    t2_headers = ["Metric", "Value", "", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, 1):
        hcell = ws.cell(row=t2_hdr_row, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center", vertical="center")

    # Metrics with formulas or placeholders
    metrics = [
        ("Test Automation Rate", "='Testing Framework'!C6"),
        ("Code Coverage", "='Test Coverage'!F6"),
        ("Rollback Success Rate", "='Rollback Procedures'!H21"),
        ("Defects Escaped to Production", "='Production Validation'!D6"),
    ]
    r = t2_hdr_row
    for metric, value in metrics:
        r += 1
        ws[f"A{r}"] = metric
        ws[f"A{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"A{r}"].border = border
        ws[f"B{r}"] = value
        ws[f"B{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"B{r}"].border = border
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border

    # TABLE 2 buffer rows (2 empty white rows)
    for _ in range(2):
        r += 1
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border

    # --- TABLE 3: CRITICAL FINDINGS ---
    crit_start = r + 2
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col_idx, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Testing Requirements", "Changes without testing", '=COUNTIF(\'Testing Framework\'!C6:C55,"❌ Non-Compliant")', "Critical", "Immediate"),
        ("Rollback Plans", "Changes without rollback plan", '=COUNTIF(\'Rollback Procedures\'!D6:D17,"❌ Non-Compliant")', "Critical", "Immediate"),
        ("Rollback Testing", "Untested rollback procedures", '=COUNTIF(\'Rollback Procedures\'!H21:H103,"❌ Non-Compliant")', "Critical", "Immediate"),
        ("Validation Methods", "Missing validation procedures", '=COUNTIF(\'Production Validation\'!D6:D62,"❌ Non-Compliant")', "Critical", "Immediate"),
        ("Testing Requirements", "Incomplete testing", '=COUNTIF(\'Testing Framework\'!C6:C55,"⚠ Partial")', "High", "Urgent"),
        ("Rollback Plans", "Incomplete rollback plans", '=COUNTIF(\'Rollback Procedures\'!D6:D17,"⚠ Partial")', "High", "Urgent"),
        ("Rollback Testing", "Partial rollback testing", '=COUNTIF(\'Rollback Procedures\'!H21:H103,"⚠ Partial")', "Medium", "Plan"),
        ("Validation Methods", "Validation method gaps", '=COUNTIF(\'Production Validation\'!D6:D62,"⚠ Partial")', "Medium", "Plan"),
    ]

    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row = crit_start + 1
    for cat, finding, formula, severity, action in findings:
        row += 1
        ws.cell(row=row, column=1, value=cat).border = border
        ws.cell(row=row, column=1).font = Font(color="000000")
        ws.cell(row=row, column=1).fill = ffffcc_fill
        ws.cell(row=row, column=2, value=finding).border = border
        ws.cell(row=row, column=2).font = Font(color="000000")
        ws.cell(row=row, column=2).fill = ffffcc_fill
        cell_count = ws.cell(row=row, column=3)
        cell_count.value = formula
        cell_count.border = border
        cell_count.alignment = Alignment(horizontal="center")
        cell_count.font = Font(color="000000")
        cell_count.fill = ffffcc_fill
        ws.cell(row=row, column=4, value=severity).border = border
        ws.cell(row=row, column=4).font = Font(color="000000")
        ws.cell(row=row, column=4).fill = ffffcc_fill
        ws.cell(row=row, column=5, value=action).border = border
        ws.cell(row=row, column=5).font = Font(color="000000")
        ws.cell(row=row, column=5).fill = ffffcc_fill
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = ffffcc_fill
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).fill = ffffcc_fill

    # TABLE 3 buffer rows (2 empty FFFFCC rows)
    for _ in range(2):
        row += 1
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border

    # Column widths (19 columns)
    gap_widths = [
        ("A", 15), ("B", 20), ("C", 40), ("D", 30), ("E", 30),
        ("F", 10), ("G", 12), ("H", 12), ("I", 12), ("J", 35),
        ("K", 20), ("L", 15), ("M", 15), ("N", 12), ("O", 15),
        ("P", 25), ("Q", 25), ("R", 30), ("S", 20)
    ]
    for col, w in gap_widths:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"


# Continue to final part with Evidence Register, Approval, and Main...
# ============================================================================
# SECTION 10: EVIDENCE_REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create standard 8-column Evidence Register with 100 rows."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Evidence type dropdown
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ev_type_dv.error = "Please select a valid evidence type"
    ev_type_dv.errorTitle = "Invalid Evidence Type"
    ws.add_data_validation(ev_type_dv)

    # Verification status dropdown
    ver_status_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Verified,{WARNING} Pending,{XMARK} Not Verified,N/A"',
        allow_blank=True,
    )
    ver_status_dv.error = "Please select a valid status"
    ver_status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(ver_status_dv)

    # Sample row (row 5) with example data (grey fill per Option B standard)
    sample_data = {
        1: "EV-001",
        2: "Testing Validation",
        3: "Test Result",
        4: "Pre-production test results for change CHG-2024-001",
        5: "\\\\fileserver\\evidence\\test_results_chg2024001.pdf",
        6: "15.01.2026",
        7: "QA Manager",
        8: f"{CHECK} Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Empty data rows (rows 6-105) - 100 empty rows for user data (MAX-002 standard)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            cell.value = None  # Empty - users choose their own evidence IDs
        # Dropdowns
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths
    for col, w in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL_SIGN_OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create standard Approval Sign-Off with 3-section approval workflow."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    # Apply borders to all merged cells in header row
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}1"].border = border

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}2"].border = border

    # --- ASSESSMENT SUMMARY ---
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    # Assessment Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.32.4 \u2014 Testing & Validation Assessment", False),
        ("Assessment Period:", "", True),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G14", False),  # TABLE 1 TOTAL row
        ("Assessed By:", "", True),
        ("Assessment Status:", "", "dropdown"),
    ]
    row = 4
    for label, value, editable in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if editable == "dropdown":
            status_dv.add(ws[f"B{row}"])
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        elif editable:
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        else:
            # Non-editable fields still need borders
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].border = border
        row += 1

    # --- Helper for approver sections ---
    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        # Apply borders to all merged cells in section title
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{start_row}"].border = border
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = border
            ws.merge_cells(f"B{r}:E{r}")
            # Apply borders and fill to all merged cells B-E
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{r}"].border = border
            r += 1
        return r

    # COMPLETED BY
    row += 1
    row = _approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")

    # REVIEWED BY
    row += 1
    row = _approver_section(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")

    # APPROVED BY (CISO)
    row += 1
    row = _approver_section(row, "APPROVED BY \u2014 CISO", "003366")

    # --- FINAL DECISION ---
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all merged cells B-E
    for col in ["B", "C", "D", "E"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # --- NEXT REVIEW DETAILS ---
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    for field in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        row += 1
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all merged cells B-E
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

    # Column widths
    for col, w in [("A", 32), ("B", 25), ("C", 20), ("D", 20), ("E", 20)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "The first principle is that you must not fool yourself
    and you are the easiest person to fool." - Richard Feynman
    
    This script generates evidence-based testing assessment tools, not cargo cult compliance.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.32.4 - Testing & Validation Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.32: Change Management")
    logger.info("Related Controls: 8.29 (Security Testing), 8.32 (Change Testing)")
    logger.info("=" * 78)
    logger.info("\n Systems Engineering Approach: Evidence-Based Compliance")
    logger.info(f" Control 8.29: Security Testing Integration")
    logger.info(f" Control 8.32: Testing Before Production Deployment")
    logger.info(f" Technology-Agnostic: Works with ANY testing framework")
    logger.info(" Audit-Ready: Comprehensive evidence collection")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("{CHECK} Workbook created with 11 sheets (per IMP specification)")

    # Create all sheets (per IMP specification - 11 sheets)
    logger.info("\n[Phase 2] Generating assessment sheets...")

    logger.info("  [1/11] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✅ Instructions complete")

    logger.info("  [2/11] Creating Testing Framework...")
    create_testing_framework_assessment(wb["Testing Framework"], styles)
    logger.info("  ✅ Testing framework assessment complete")

    logger.info("  [3/11] Creating Test Coverage...")
    create_test_types_coverage(wb["Test Coverage"], styles)
    logger.info("  ✅ Test coverage assessment complete")

    logger.info("  [4/11] Creating Acceptance Criteria...")
    create_acceptance_criteria_management(wb["Acceptance Criteria"], styles)
    logger.info("  ✅ Acceptance criteria management complete")

    logger.info("  [5/11] Creating Rollback Procedures...")
    create_rollback_procedures(wb["Rollback Procedures"], styles)
    logger.info("  ✅ Rollback procedures assessment complete")

    logger.info("  [6/11] Creating Production Validation...")
    create_production_validation(wb["Production Validation"], styles)
    logger.info("  ✅ Production validation assessment complete")

    logger.info("  [7/11] Creating Security Testing...")
    create_security_testing(wb["Security Testing"], styles)
    logger.info("  ✅ Security testing assessment complete")

    logger.info("  [8/11] Creating Testing Documentation...")
    create_testing_documentation(wb["Testing Documentation"], styles)
    logger.info("  ✅ Testing documentation complete")

    logger.info("  [9/11] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  ✅ Evidence register complete (100 evidence rows)")

    logger.info("  [10/11] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  ✅ Summary dashboard complete")

    logger.info("  [11/11] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  ✅ Approval workflow complete")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.32.4_Testing_Validation_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"{XMARK} ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info(" WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets:")
    logger.info("  • Instructions & Legend (testing principles, Control 8.29 & 8.32 guidance)")
    logger.info("  • Testing Framework (15 strategy aspects, 15 tools, 9 metrics)")
    logger.info("  • Test Coverage (~80 test aspects across 6 test types)")
    logger.info("    - Unit Testing (11 aspects)")
    logger.info("    - Integration Testing (11 aspects)")
    logger.info("    - System/Functional Testing (11 aspects)")
    logger.info("    - UAT Assessment (10 aspects)")
    logger.info("    - Security Testing - Control 8.29 (20 aspects)")
    logger.info("    - Performance Testing (17 aspects)")
    logger.info("  • Acceptance Criteria (13 framework aspects, 12 change types, 12 stages)")
    logger.info("  • Rollback Procedures (12 strategy aspects, 18 change types, 20 test tracking rows)")
    logger.info("  • Production Validation (12 strategy aspects, 35+ validation checks)")
    logger.info("\n Analysis & Governance:")
    logger.info("  • Summary Dashboard (7 assessment areas, SUM formulas, key metrics)")
    logger.info("  • Evidence Register (100 evidence entries)")
    logger.info("  • Approval Sign-Off (3-section approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info(" ASSESSMENT CAPABILITIES:")
    logger.info("  • Comprehensive testing framework assessment")
    logger.info("  • Test type coverage (unit, integration, UAT, security, performance)")
    logger.info("  • Security testing integration (Control 8.29)")
    logger.info("  • Acceptance criteria definition and tracking")
    logger.info("  • Rollback capability assessment by change type")
    logger.info("  • Production validation procedures")
    logger.info("  • 100 evidence documentation entries")
    logger.info("  • Automated compliance calculations")
    logger.info("  • Test metrics tracking")
    logger.info("\n" + "─" * 78)
    logger.info(f" KEY FEATURES:")
    logger.info("  ✅ Technology-agnostic (works with ANY testing framework)")
    logger.info("  ✅ Control 8.29 (Security Testing) comprehensive assessment")
    logger.info("  ✅ Control 8.32 (Change Testing) complete coverage")
    logger.info("  ✅ Test automation assessment")
    logger.info("  ✅ Rollback testing verification")
    logger.info("  ✅ Production validation procedures")
    logger.info("  ✅ Comprehensive evidence collection")
    logger.info("  ✅ Multi-level approval workflow")
    logger.info("  ✅ Quarterly review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(f" NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Complete Instructions & Legend sheet first")
    logger.info("  3. Document testing framework and tools")
    logger.info("  4. Assess coverage of all test types (unit, integration, security, etc.)")
    logger.info("  5. Define acceptance criteria for each change type")
    logger.info("  6. Document rollback procedures and testing")
    logger.info("  7. Define production validation procedures")
    logger.info("  8. Review Summary Dashboard for compliance metrics")
    logger.info("  9. Document evidence in Evidence Register")
    logger.info("  10. Obtain final approval via Approval Sign-Off")
    logger.info("\n PRO TIP:")
    logger.info("  Control 8.29 is CRITICAL: Security testing must be integrated into")
    logger.info("  your development and acceptance processes. Don't treat security testing")
    logger.info("  as an afterthought - shift left and test early!")
    logger.info("\n" + "=" * 78)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info("\n This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
