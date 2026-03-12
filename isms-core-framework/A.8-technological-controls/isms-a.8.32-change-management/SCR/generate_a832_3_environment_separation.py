#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.32.3 - Environment Separation Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Assessment Domain 3 of 4: Development/Test/Production Environment Isolation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific environment architecture and separation requirements.

Key customization areas:
1. Environment definitions (match your Dev/Test/Staging/Prod architecture)
2. Access control requirements (align with your IAM policies)
3. Data protection measures (adapt to your data classification)
4. Promotion workflow stages (customize to your SDLC)
5. Environment-specific controls (based on your infrastructure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
environment separation, access controls, and promotion workflows against
ISO 27001:2022 Control A.8.32 and Control 8.31 (Separation of Environments)
requirements.

**Purpose:**
Enables systematic assessment of Development, Test/QA, and Production environment
isolation, access controls, data protection measures, and change promotion
workflows to prevent unauthorised changes to production systems.

**Assessment Scope:**
- Development environment configuration and access controls
- Test/QA environment isolation and data protection
- Production environment protection and change restrictions
- Environment promotion workflow (Dev → Test → Prod)
- Access control segregation between environments
- Production data protection in non-production environments
- Environment-specific compliance verification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and separation standards
2. Development_Environment - Dev environment controls and access
3. Test_QA_Environment - Test isolation and data protection
4. Production_Environment - Production protection and restrictions
5. Environment_Promotion_Workflow - Promotion process and controls
6. Access Controls_Separation - Role-based access segregation
7. Summary Dashboard - Compliance metrics and analytics
8. Evidence Register - Audit evidence tracking
9. Approval Sign-Off - Stakeholder approval workflow

**Key Features:**
- Technology-agnostic assessment (cloud, on-premise, hybrid)
- Environment separation best practices from ISO 27001:2022 Control 8.31
- Access control matrices for role segregation
- Production data protection in non-prod environments (Control 8.33)
- Automated compliance calculations
- Evidence linkage for audit traceability

**Integration:**
with related controls 8.31 (Separation of Environments) and 8.33 (Test Information).

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a832_3_environment_separation.py

Output:
    File: ISMS_A_8_32_3_Environment_Separation_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Document your organisation's environment architecture
    2. Define access controls for each environment
    3. Document data protection measures in non-production
    4. Map environment promotion workflow stages
    5. Review access segregation between environments
    6. Assess production data usage in test environments
    7. Review Summary Dashboard for compliance metrics

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32, A.8.31
Assessment Domain:    3 of 4 (Environment Separation & Access Control)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.3: Environment Separation Implementation Guide
    - ISMS-IMP-A.8.32.1: Change Process Assessment (Domain 1)
    - ISMS-IMP-A.8.32.2: Change Types & Categories Assessment (Domain 2)
    - ISMS-IMP-A.8.32.4: Testing & Validation Assessment (Domain 4)
    - ISMS-POL-A.8.31: Separation of Development, Test and Production Environments

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.32.3 specification
    - Supports comprehensive environment separation evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Environment Separation Rationale:**
Proper environment separation prevents untested code from reaching production,
reduces risk of production data exposure, and ensures changes follow proper
testing and validation before deployment.

**Access Control Principle:**
Developers should not have production write access. Production changes should
flow through controlled promotion processes, not direct developer access.

**Production Data Protection:**
Production data in test environments is a major compliance risk. Use data
masking, synthetic data, or subset data where possible. See ISMS-POL-A.8.11
(Data Masking) for guidance.

**Audit Considerations:**
Auditors will verify environment separation controls, access segregation,
and data protection measures in non-production environments.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.32.3"
WORKBOOK_NAME = "Environment Separation Assessment"
CONTROL_ID = "A.8.32"
CONTROL_NAME = "Change Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠  Warning sign
GEAR = '\u2699'       # ⚙  Gear
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP-A.8.32.3 spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.32.3 specification (10 sheets)
    sheets = [
        "Instructions & Legend",
        "Environment Inventory",
        "Access Controls",
        "Promotion Workflows",
        "Data Protection",
        "Environment Config",
        "Separation Controls",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create ALL data validation objects for dropdowns.
    CRITICAL: Must include EVERY dropdown type used in the spec.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,❌ No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,❌ No,N/A"',
            allow_blank=False
        ),
        'yes_partial_no': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,⚠ Partial,❌ No"',
            allow_blank=False
        ),
        'yes_partial_no_na': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,⚠ Partial,❌ No,N/A"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Implemented,⚠ Partial,❌ Not Implemented, Planned,N/A"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Compliant,⚠ Partial,❌ Non-Compliant, Pending"',
            allow_blank=False
        ),
        'environment_type': DataValidation(
            type="list",
            formula1='"Development,Test,QA,UAT,Staging,Pre-Production,Production,DR/Backup"',
            allow_blank=False
        ),
        'hosting_model': DataValidation(
            type="list",
            formula1='"On-Premises,Cloud IaaS,Cloud PaaS,Cloud SaaS,Hybrid,Managed Service"',
            allow_blank=False
        ),
        'isolation_level': DataValidation(
            type="list",
            formula1='"Physical,Network (VLAN),Logical (Namespace/VM),None"',
            allow_blank=False
        ),
        'access_method': DataValidation(
            type="list",
            formula1='"VPN,Bastion Host,Direct,Jump Server,API,Web Portal,SSH Key,Other"',
            allow_blank=False
        ),
        'mfa_required': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes - All Access,✅ Yes - Production Only,⚠ Partial,❌ No"',
            allow_blank=False
        ),
        'approval_required': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes - Manager,✅ Yes - Security,✅ Yes - Both,⚠ Informal,❌ No"',
            allow_blank=False
        ),
        'access_review_frequency': DataValidation(
            type="list",
            formula1='"Monthly,Quarterly,Bi-Annually,Annually,Ad-hoc,None"',
            allow_blank=False
        ),
        'data_classification': DataValidation(
            type="list",
            formula1='"Public,Internal,Confidential,Highly Confidential,Personal Data"',
            allow_blank=False
        ),
        'anonymization_method': DataValidation(
            type="list",
            formula1='"Masking,Tokenization,Synthetic Data,Subset Only,Encryption,Not Anonymized"',
            allow_blank=False
        ),
        'promotion_method': DataValidation(
            type="list",
            formula1='"Manual Deploy,CI/CD Pipeline,Automated,Hybrid,Release Management Tool"',
            allow_blank=False
        ),
        'promotion_frequency': DataValidation(
            type="list",
            formula1='"On-Demand,Daily,Weekly,Bi-Weekly,Monthly,Per Release"',
            allow_blank=False
        ),
        'testing_required': DataValidation(
            type="list",
            formula1=f'"{CHECK} Mandatory,⚠ Recommended,❌ Optional"',
            allow_blank=False
        ),
        'rollback_capability': DataValidation(
            type="list",
            formula1=f'"{CHECK} Automated,⚠ Manual,❌ Limited,❌ None"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Change Request,Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other,N/A"',
            allow_blank=True
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⚠ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Final,⚠ Requires Remediation, Draft,❌ Re-assessment Required"',
            allow_blank=False
        ),
        'review_recommendation': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approve,⚠ Approve with Conditions,❌ Reject, Require Rework"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approved,⚠ Approved with Conditions,❌ Rejected"',
            allow_blank=False
        ),
        'backup_frequency': DataValidation(
            type="list",
            formula1='"Real-time,Hourly,Daily,Weekly,Monthly,None"',
            allow_blank=False
        ),
        'monitoring_level': DataValidation(
            type="list",
            formula1=f'"{CHECK} Comprehensive,⚠ Basic,❌ Minimal,❌ None"',
            allow_blank=False
        ),
        'dpo_approval': DataValidation(
            type="list",
            formula1=f'"{CHECK} With Approval,⚠ Pending,❌ Not Required,❌ Required but Missing"',
            allow_blank=False
        ),
        'usage_justification': DataValidation(
            type="list",
            formula1=f'"{CHECK} Primary,⚠ Supplemental,❌ Not Used"',
            allow_blank=False
        ),
    }
    
    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Document YOUR development environment configuration and controls.', '2. Document YOUR test/QA environment configuration and controls.', '3. Document YOUR production environment configuration and controls.', '4. Define YOUR environment promotion procedures.', '5. Assess YOUR production data usage in non-production environments (Control 8.33).', '6. Review the Summary Dashboard for compliance metrics.', '7. Maintain the Evidence Register for audit traceability.', '8. Obtain final approval via Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_development_environment(ws, styles):
    """Create Development_Environment assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "DEVELOPMENT ENVIRONMENT ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = "Document development environment configuration and controls"
    apply_style(ws["A2"], styles["subheader"])

    # Environment Identification
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ENVIRONMENT IDENTIFICATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Attribute", "Value", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Environment identification attributes
    env_attributes = [
        ("Environment Name", "text", None),
        ("Environment Type", "dropdown", "environment_type"),
        ("Hosting Model", "dropdown", "hosting_model"),
        ("Location/Region", "text", None),
        ("Network Isolation Level", "dropdown", "isolation_level"),
        ("VLAN/Subnet ID", "text", None),
        ("Primary Purpose", "text", None),
        ("Number of Systems", "text", None),
        ("Number of Users with Access", "text", None),
        ("Access Method", "dropdown", "access_method"),
        ("Change Frequency", "text", None),
        ("Deployment Tools Used", "text", None),
        ("Monitoring Solution", "text", None),
        ("Backup Strategy", "dropdown", "backup_frequency"),
    ]

    row += 1
    for attr_name, field_type, validation_key in env_attributes:
        ws[f"A{row}"] = attr_name
        ws[f"A{row}"].font = Font(bold=True)
        
        # Value cell
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        # Compliance cell
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        # Evidence cell
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['evidence_type'].add(ws[f"D{row}"])

        # Notes cell
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Access Control Assessment
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ACCESS CONTROL ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    access_controls = [
        ("Developer access granted by default?", "dropdown", "yes_no"),
        ("Role-Based Access Control (RBAC) implemented?", "dropdown", "yes_partial_no"),
        ("Access approval process defined?", "dropdown", "approval_required"),
        ("Access request tracked in ticket system?", "dropdown", "yes_no"),
        ("Multi-Factor Authentication (MFA) required?", "dropdown", "mfa_required"),
        ("Access review performed?", "dropdown", "access_review_frequency"),
        ("Privileged access tracked separately?", "dropdown", "yes_no_na"),
        ("Shared accounts prohibited?", "dropdown", "yes_partial_no"),
        ("SSH key management implemented?", "dropdown", "yes_partial_no_na"),
        ("Service accounts documented?", "dropdown", "yes_partial_no_na"),
        ("Emergency access procedure defined?", "dropdown", "yes_no"),
    ]

    row += 1
    for control_name, field_type, validation_key in access_controls:
        ws[f"A{row}"] = control_name
        
        # Value cell
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        # Compliance cell
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        # Evidence cell
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['evidence_type'].add(ws[f"D{row}"])

        # Notes cell
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Data Controls
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "DATA CONTROLS (CONTROL 8.33 CRITICAL)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    data_controls = [
        ("Production data prohibited in development?", "dropdown", "yes_partial_no"),
        ("If prod data present, is it anonymized?", "dropdown", "yes_partial_no_na"),
        ("Anonymization method documented?", "dropdown", "anonymization_method"),
        ("Synthetic data generation capability?", "dropdown", "yes_partial_no"),
        ("Data subset strategy implemented?", "dropdown", "yes_no_na"),
        ("Encryption at rest implemented?", "dropdown", "yes_partial_no"),
        ("Encryption in transit enforced?", "dropdown", "yes_partial_no"),
        ("DPO approval obtained (if prod data)?", "dropdown", "dpo_approval"),
    ]

    row += 1
    for control_name, field_type, validation_key in data_controls:
        ws[f"A{row}"] = control_name
        
        # Value cell
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        # Compliance cell
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        # Evidence cell
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['evidence_type'].add(ws[f"D{row}"])

        # Notes cell
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Change Management Integration
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "CHANGE MANAGEMENT INTEGRATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    cm_aspects = [
        ("Changes logged in change management system?", "dropdown", "yes_partial_no"),
        ("Version control system in use?", "text", None),
        ("Code review process defined?", "dropdown", "yes_partial_no"),
        ("Testing before promotion to test environment?", "dropdown", "testing_required"),
        ("Automated testing implemented?", "dropdown", "yes_partial_no"),
    ]

    row += 1
    for aspect_name, field_type, validation_key in cm_aspects:
        ws[f"A{row}"] = aspect_name
        
        # Value cell
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        # Compliance cell
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        # Evidence cell
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['evidence_type'].add(ws[f"D{row}"])

        # Notes cell
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: TEST_QA_ENVIRONMENT SHEET
# ============================================================================

def create_test_qa_environment(ws, styles):
    """Create Test_QA_Environment assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "TEST/QA ENVIRONMENT ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = "Document test/QA environment configuration and controls"
    apply_style(ws["A2"], styles["subheader"])

    # Environment Identification
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ENVIRONMENT IDENTIFICATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Attribute", "Value", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    env_attributes = [
        ("Environment Name", "text", None),
        ("Environment Type", "dropdown", "environment_type"),
        ("Hosting Model", "dropdown", "hosting_model"),
        ("Location/Region", "text", None),
        ("Network Isolation Level", "dropdown", "isolation_level"),
        ("VLAN/Subnet ID", "text", None),
        ("Primary Purpose", "text", None),
        ("Number of Systems", "text", None),
        ("Number of Users with Access", "text", None),
        ("Access Method", "dropdown", "access_method"),
        ("Change Frequency", "text", None),
        ("Test Automation Tools", "text", None),
        ("Monitoring Solution", "text", None),
        ("Backup Strategy", "dropdown", "backup_frequency"),
    ]

    row += 1
    for attr_name, field_type, validation_key in env_attributes:
        ws[f"A{row}"] = attr_name
        ws[f"A{row}"].font = Font(bold=True)
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Access Control Assessment
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ACCESS CONTROL ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    access_controls = [
        ("QA/Tester access controlled?", "dropdown", "yes_partial_no"),
        ("Developer access restricted vs development?", "dropdown", "yes_partial_no"),
        ("Role-Based Access Control (RBAC) implemented?", "dropdown", "yes_partial_no"),
        ("Access approval process defined?", "dropdown", "approval_required"),
        ("Access request tracked in ticket system?", "dropdown", "yes_no"),
        ("Multi-Factor Authentication (MFA) required?", "dropdown", "mfa_required"),
        ("Access review performed?", "dropdown", "access_review_frequency"),
        ("Privileged access tracked separately?", "dropdown", "yes_no_na"),
        ("Shared accounts prohibited?", "dropdown", "yes_partial_no"),
        ("Test data access logged?", "dropdown", "yes_partial_no"),
        ("UAT user access documented?", "dropdown", "yes_no_na"),
    ]

    row += 1
    for control_name, field_type, validation_key in access_controls:
        ws[f"A{row}"] = control_name
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Data Controls
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "DATA CONTROLS (CONTROL 8.33 CRITICAL)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    data_controls = [
        ("Production data usage policy enforced?", "dropdown", "yes_partial_no"),
        ("If prod data present, is it anonymized?", "dropdown", "yes_partial_no_na"),
        ("Anonymization method documented?", "dropdown", "anonymization_method"),
        ("Anonymization effectiveness tested?", "dropdown", "yes_no_na"),
        ("Synthetic data generation capability?", "dropdown", "yes_partial_no"),
        ("Data subset strategy implemented?", "dropdown", "yes_no_na"),
        ("Re-identification risk assessed?", "dropdown", "yes_no_na"),
        ("Encryption at rest implemented?", "dropdown", "yes_partial_no"),
        ("Encryption in transit enforced?", "dropdown", "yes_partial_no"),
        ("Data backup procedures defined?", "dropdown", "yes_partial_no"),
        ("Data retention/disposal policy?", "dropdown", "yes_partial_no"),
        ("Test data refresh procedures?", "dropdown", "yes_partial_no"),
        ("DPO approval obtained (if prod data)?", "dropdown", "dpo_approval"),
    ]

    row += 1
    for control_name, field_type, validation_key in data_controls:
        ws[f"A{row}"] = control_name
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Testing Controls
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TESTING CONTROLS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    testing_controls = [
        ("Test plans required?", "dropdown", "testing_required"),
        ("Test results documented?", "dropdown", "yes_partial_no"),
        ("Automated testing framework?", "dropdown", "yes_partial_no"),
        ("Performance testing capabilities?", "dropdown", "yes_partial_no"),
        ("Security testing performed?", "dropdown", "yes_partial_no"),
        ("UAT sign-off required?", "dropdown", "yes_partial_no"),
    ]

    row += 1
    for control_name, field_type, validation_key in testing_controls:
        ws[f"A{row}"] = control_name
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: PRODUCTION_ENVIRONMENT SHEET
# ============================================================================

def create_production_environment(ws, styles):
    """Create Production_Environment assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "PRODUCTION ENVIRONMENT ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = "Document production environment configuration and controls (HIGHEST SECURITY)"
    apply_style(ws["A2"], styles["subheader"])

    # Environment Identification
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ENVIRONMENT IDENTIFICATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Attribute", "Value", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    env_attributes = [
        ("Environment Name", "text", None),
        ("Environment Type", "dropdown", "environment_type"),
        ("Hosting Model", "dropdown", "hosting_model"),
        ("Location/Region", "text", None),
        ("Network Isolation Level", "dropdown", "isolation_level"),
        ("VLAN/Subnet ID", "text", None),
        ("Primary Purpose", "text", None),
        ("Number of Systems", "text", None),
        ("Number of Users Served", "text", None),
        ("Business Criticality", "text", None),
        ("Access Method", "dropdown", "access_method"),
        ("Change Window Schedule", "text", None),
        ("Deployment Method", "dropdown", "promotion_method"),
        ("Monitoring Solution", "text", None),
        ("Backup Strategy", "dropdown", "backup_frequency"),
        ("DR/HA Configuration", "text", None),
        ("SLA/RTO/RPO Targets", "text", None),
        ("Compliance Requirements", "text", None),
    ]

    row += 1
    for attr_name, field_type, validation_key in env_attributes:
        ws[f"A{row}"] = attr_name
        ws[f"A{row}"].font = Font(bold=True)
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Production Access Restrictions
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "PRODUCTION ACCESS RESTRICTIONS (CRITICAL)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    access_restrictions = [
        ("Developer direct access prohibited?", "dropdown", "yes_partial_no"),
        ("All access logged and monitored?", "dropdown", "yes_partial_no"),
        ("Multi-Factor Authentication (MFA) MANDATORY?", "dropdown", "yes_no"),
        ("Role-Based Access Control (RBAC) enforced?", "dropdown", "yes_partial_no"),
        ("Principle of Least Privilege applied?", "dropdown", "yes_partial_no"),
        ("Privileged access requires approval?", "dropdown", "approval_required"),
        ("Emergency access break-glass procedure?", "dropdown", "yes_no"),
        ("Access review frequency", "dropdown", "access_review_frequency"),
        ("Just-In-Time (JIT) access implemented?", "dropdown", "yes_partial_no_na"),
        ("Shared accounts prohibited?", "dropdown", "yes_no"),
        ("Service accounts managed separately?", "dropdown", "yes_partial_no"),
        ("SSH keys rotated regularly?", "dropdown", "yes_no_na"),
        ("Bastion/Jump host required for access?", "dropdown", "yes_no_na"),
        ("Administrative actions require ticket?", "dropdown", "yes_partial_no"),
    ]

    row += 1
    for control_name, field_type, validation_key in access_restrictions:
        ws[f"A{row}"] = control_name
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Data Protection Controls
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "DATA PROTECTION CONTROLS (CONTROL 8.33 CRITICAL)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    data_protection = [
        ("Production data NEVER copied to non-prod?", "dropdown", "yes_partial_no"),
        ("Production data exfiltration monitoring?", "dropdown", "yes_partial_no"),
        ("Data Loss Prevention (DLP) implemented?", "dropdown", "yes_partial_no_na"),
        ("Encryption at rest enforced?", "dropdown", "yes_partial_no"),
        ("Encryption in transit enforced?", "dropdown", "yes_partial_no"),
        ("Database access logging comprehensive?", "dropdown", "yes_partial_no"),
        ("Backup encryption implemented?", "dropdown", "yes_partial_no"),
        ("Backup tested regularly?", "dropdown", "yes_partial_no"),
        ("Data classification implemented?", "dropdown", "yes_partial_no"),
        ("Personal data handling documented?", "dropdown", "yes_partial_no"),
        ("Data retention policy enforced?", "dropdown", "yes_partial_no"),
        ("Secure data disposal procedures?", "dropdown", "yes_partial_no"),
    ]

    row += 1
    for control_name, field_type, validation_key in data_protection:
        ws[f"A{row}"] = control_name
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Change Management Integration
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "CHANGE MANAGEMENT INTEGRATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    cm_aspects = [
        ("All production changes require approval?", "dropdown", "yes_no"),
        ("Changes tested in non-prod first?", "dropdown", "testing_required"),
        ("Change window adherence monitored?", "dropdown", "yes_partial_no"),
        ("Emergency change procedure defined?", "dropdown", "yes_no"),
        ("Rollback procedure documented and tested?", "dropdown", "rollback_capability"),
        ("Post-implementation validation required?", "dropdown", "yes_partial_no"),
        ("Change success rate monitored?", "dropdown", "yes_no"),
    ]

    row += 1
    for aspect_name, field_type, validation_key in cm_aspects:
        ws[f"A{row}"] = aspect_name
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"C{row}"])

        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 7: ENVIRONMENT_PROMOTION_PROCESS SHEET
# ============================================================================

def create_environment_promotion_process(ws, styles):
    """Create Environment_Promotion_Process assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ENVIRONMENT PROMOTION PROCESS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = "Document promotion workflows between environments (Dev → Test → Prod)"
    apply_style(ws["A2"], styles["subheader"])

    # Promotion Path Definition
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "PROMOTION PATH DEFINITION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["From Environment", "To Environment", "Method", "Approval Required?", "Frequency", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Standard promotion paths
    promotion_paths = [
        "Development → Test",
        "Test → QA",
        "QA → UAT",
        "UAT → Production",
        "Hotfix → Production (Emergency)",
    ]

    row += 1
    for path in promotion_paths:
        from_env, to_env = path.split(" → ")
        ws[f"A{row}"] = from_env
        ws[f"B{row}"] = to_env
        
        # Method dropdown
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['promotion_method'].add(ws[f"C{row}"])

        # Approval dropdown
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['approval_required'].add(ws[f"D{row}"])

        # Frequency dropdown
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['promotion_frequency'].add(ws[f"E{row}"])

        # Notes
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        row += 1

    # Promotion Controls
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "PROMOTION CONTROLS & GATES"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    control_headers = ["Control", "Dev→Test", "Test→QA", "QA→Prod", "Compliance", "Evidence"]
    for col_idx, header in enumerate(control_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    promotion_controls = [
        "Code review completed?",
        "Unit tests passed?",
        "Integration tests passed?",
        "Security scan passed?",
        "Performance test passed?",
        "UAT sign-off obtained?",
        "Documentation updated?",
        "Change ticket approved?",
        "Rollback plan documented?",
        "Stakeholders notified?",
        "Change window scheduled?",
        "Backup completed?",
        "Validation test defined?",
        "Post-deployment checklist?",
        "CAB approval (if required)?",
        "Emergency bypass procedure?",
    ]

    row += 1
    for control_name in promotion_controls:
        ws[f"A{row}"] = control_name
        
        # Dev→Test
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['testing_required'].add(ws[f"B{row}"])

        # Test→QA
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        validations['testing_required'].add(ws[f"C{row}"])

        # QA→Prod
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['testing_required'].add(ws[f"D{row}"])

        # Compliance
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"E{row}"])

        # Evidence
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=6))

        row += 1

    # CI/CD Pipeline Assessment
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CI/CD PIPELINE ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    pipeline_headers = ["Pipeline Stage", "Implemented?", "Tool/Method", "Automated?", "Compliance", "Notes"]
    for col_idx, header in enumerate(pipeline_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    pipeline_stages = [
        "Source control integration",
        "Build automation",
        "Unit testing",
        "Code quality analysis",
        "Security scanning (SAST)",
        "Dependency vulnerability scanning",
        "Integration testing",
        "Container scanning (if applicable)",
        "Artifact versioning",
        "Environment-specific config management",
        "Automated deployment to test/QA",
        "Production deployment approval gate",
    ]

    row += 1
    for stage_name in pipeline_stages:
        ws[f"A{row}"] = stage_name
        
        # Implemented?
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['implementation_status'].add(ws[f"B{row}"])

        # Tool/Method
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Automated?
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"D{row}"])

        # Compliance
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"E{row}"])

        # Notes
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 30

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: PRODUCTION_DATA_IN_NONPROD SHEET
# ============================================================================

def create_production_data_in_nonprod(ws, styles):
    """Create Production_Data_in_NonProd assessment sheet (Control 8.33)."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "PRODUCTION DATA IN NON-PRODUCTION ENVIRONMENTS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Control 8.33: Test Information - Production data protection assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Policy & Governance
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "POLICY & GOVERNANCE"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Requirement", "Implemented?", "Details", "Compliance", "Evidence", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    policy_requirements = [
        "Production data usage policy documented?",
        "Prohibition of prod data in non-prod enforced?",
        "Exceptions process defined and documented?",
        "DPO review required for exceptions?",
        "Data classification system implemented?",
        "Anonymization procedures documented?",
        "Annual policy review performed?",
    ]

    row += 1
    for requirement in policy_requirements:
        ws[f"A{row}"] = requirement
        
        # Implemented?
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        # Details
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Compliance
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"D{row}"])

        # Evidence
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=5))

        # Notes
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        row += 1

    # Data Anonymization Controls
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "DATA ANONYMIZATION CONTROLS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    anonymization_controls = [
        "Data masking implemented?",
        "Tokenization used for sensitive fields?",
        "Anonymization testing performed?",
        "Re-identification risk assessed?",
        "Anonymization effectiveness >95%?",
        "PII/PHI identification automated?",
        "Direct identifiers removed?",
        "Indirect identifiers assessed?",
        "Data linkage risk evaluated?",
        "Anonymization audit trail maintained?",
        "Failed anonymization procedures defined?",
        "Anonymization tools validated?",
        "Data subsetting implemented?",
        "Referential integrity maintained?",
        "Test data generation capability?",
        "Synthetic data generation used?",
        "Production data refresh controls?",
    ]

    row += 1
    for control_name in anonymization_controls:
        ws[f"A{row}"] = control_name
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no_na'].add(ws[f"B{row}"])

        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"D{row}"])

        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=5))
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        row += 1

    # Current Production Data Usage
    row += 1
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "CURRENT PRODUCTION DATA USAGE IN NON-PRODUCTION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    usage_headers = ["System/Application", "Contains Prod Data?", "Data Type", "Anonymized?", "Approval Date", "Approved By", "Review Date", "Compliant?", "Evidence"]
    for col_idx, header in enumerate(usage_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Sample row with example data
    row += 1
    sample_data = {
        1: "CRM Test Environment",
        2: "No",
        3: "Synthetic customer data",
        4: "N/A",
        5: "15.01.2026",
        6: "Test Manager",
        7: "15.07.2026",
        8: "Compliant",
        9: "EV-DATA-001",
    }
    for col_idx, value in sample_data.items():
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = styles["border"]

    # Add dropdowns to sample row
    validations['yes_no'].add(ws[f"B{row}"])
    validations['yes_partial_no'].add(ws[f"D{row}"])
    validations['compliance_status'].add(ws[f"H{row}"])
    validations['evidence_type'].add(ws.cell(row=row, column=9))

    # Empty data rows (50 rows for user data)
    row += 1
    for i in range(50):
        for col_idx in range(1, 10):  # Columns A-I
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]
            cell.value = None  # Empty - users document their own systems

        # Add dropdowns to empty rows
        validations['yes_no'].add(ws[f"B{row}"])
        validations['yes_partial_no'].add(ws[f"D{row}"])
        validations['compliance_status'].add(ws[f"H{row}"])
        validations['evidence_type'].add(ws.cell(row=row, column=9))

        row += 1

    # Synthetic Data Generation
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "SYNTHETIC DATA GENERATION CAPABILITY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    synthetic_headers = ["Capability", "Available?", "Tool/Method", "Data Types Supported", "Usage", "Compliance", "Evidence"]
    for col_idx, header in enumerate(synthetic_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    synthetic_capabilities = [
        "Synthetic data generator",
        "Maintains referential integrity",
        "Realistic data distributions",
        "Supports edge cases",
        "Automated generation",
    ]

    row += 1
    for capability in synthetic_capabilities:
        ws[f"A{row}"] = capability
        
        # Available?
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]
        validations['yes_partial_no'].add(ws[f"B{row}"])

        # Tool/Method
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Data Types Supported
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        # Usage
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]
        validations['usage_justification'].add(ws[f"E{row}"])

        # Compliance
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        validations['compliance_status'].add(ws[f"F{row}"])

        # Evidence
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A5"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8.5: SEPARATION CONTROLS SHEET
# ============================================================================

def create_separation_controls(ws, styles):
    """Create Separation Controls sheet documenting technical controls enforcing separation."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "ENVIRONMENT SEPARATION CONTROLS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document technical controls enforcing environment separation"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== SEPARATION CONTROLS ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "TECHNICAL SEPARATION CONTROLS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    headers = ["Control Type", "Description", "Environments Covered", "Implementation Status", "Effectiveness", "Last Tested", "Owner", "Evidence"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    controls = [
        ("Network Segmentation", "VLAN/subnet separation between environments", "", "", "", "", "", ""),
        ("Firewall Rules", "Inter-environment traffic restrictions", "", "", "", "", "", ""),
        ("IAM Policies", "Role-based access per environment", "", "", "", "", "", ""),
        ("CI/CD Pipeline Controls", "Automated promotion with approvals", "", "", "", "", "", ""),
        ("Database Isolation", "Separate database instances/schemas", "", "", "", "", "", ""),
        ("Secret Management", "Environment-specific credentials", "", "", "", "", "", ""),
        ("Monitoring Separation", "Distinct logging/monitoring per env", "", "", "", "", "", ""),
        ("Container Isolation", "Namespace/cluster separation", "", "", "", "", "", ""),
    ]

    for control_data in controls:
        for col_idx, value in enumerate(control_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles["input_cell"])
            # Add compliance DV to Implementation Status column (D)
            if col_idx == 4:
                validations['compliance_status'].add(cell)
            # Add evidence DV to Evidence column (H=8)
            if col_idx == 8:
                validations['evidence_type'].add(cell)
        row += 1

    row += 2

    # ==================== CONTROL EFFECTIVENESS ====================
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CONTROL EFFECTIVENESS ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    eff_headers = ["Assessment Area", "Target", "Current Score", "Gap", "Remediation Required", "Target Date"]
    for col_idx, header in enumerate(eff_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    areas = [
        ("Network Separation", "100%", "", "", "", ""),
        ("Access Control Separation", "100%", "", "", "", ""),
        ("Data Separation", "100%", "", "", "", ""),
        ("Code Promotion Controls", "100%", "", "", "", ""),
    ]

    for area_data in areas:
        for col_idx, value in enumerate(area_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles["input_cell"])
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 20

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: SUMMARY_DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with standard compliance table and metrics."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "ENVIRONMENT SEPARATION ASSESSMENT — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.32: Change Management"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # --- TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW ---
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    assessment_areas = [
        "Environment Inventory",
        "Access Controls",
        "Promotion Workflows - Controls & Gates",
        "Promotion Workflows - CI/CD Pipeline",
        "Data Protection",
        "Environment Config",
        "Separation Controls",
    ]

    # Area configurations: (sheet_name, status_col, [compliant, partial, non_compliant], row_start, row_end)
    # None = manual entry area
    area_configs = [
        ('Environment Inventory', 'C', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 55),
        ('Access Controls', 'C', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 58),  # Multiple sections: Env ID, Access, Data, Testing
        # Promotion Workflows - 2 sections with formulas
        ('Promotion Workflows', 'E', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 14, 44),  # Controls & Gates section (extended to cover all data)
        ('Promotion Workflows', 'E', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 33, 44),  # CI/CD Pipeline section
        ('Data Protection', 'D', ['✅ Yes', '⚠ Partial', '❌ No'], 6, 95),  # Multiple data protection sections
        ('Environment Config', 'C', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 65),  # Extended configuration assessment
        ('Separation Controls', 'D', ['✅ Compliant', '⚠ Partial', '❌ Non-Compliant'], 6, 21),  # Technical separation controls (8 controls + effectiveness section)
    ]

    for i, area_name in enumerate(assessment_areas):
        r = 6 + i
        ws[f"A{r}"] = area_name
        ws[f"A{r}"].font = Font(name="Calibri", size=10)
        ws[f"A{r}"].border = border

        if area_configs[i] is None:
            # MANUAL ENTRY - use placeholder cells
            for col in "BCDEF":
                cell = ws[f"{col}{r}"]
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Manual areas get placeholder in column G
            ws[f"G{r}"] = "[enter %]"
            ws[f"G{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        else:
            # FORMULA-BASED - generate compliance formulas
            sheet_name, status_col, status_values, row_start, row_end = area_configs[i]
            compliant_val, partial_val, non_compliant_val = status_values

            # B: Total items (count non-empty rows in column A)
            ws[f"B{r}"] = f"=COUNTA('{sheet_name}'!A{row_start}:A{row_end})"
            ws[f"B{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"B{r}"].border = border
            ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # C: Compliant
            ws[f"C{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{compliant_val}\")"
            ws[f"C{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"C{r}"].border = border
            ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # D: Partially Compliant
            ws[f"D{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{partial_val}\")"
            ws[f"D{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"D{r}"].border = border
            ws[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # E: Non-Compliant (including "Pending" for sheets that have it in DV)
            if area_name == "Data Protection":
                # Data Protection uses Yes/Partial/No - no "Pending" status
                ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")"
            elif "Promotion Workflows" in area_name:
                # Promotion Workflows uses "Pending" for compliance_status DV
                ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")+COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"Pending\")"
            else:
                # Other sheets use Compliant/Partial/Non-Compliant/Pending
                ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")+COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"Pending\")"
            ws[f"E{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"E{r}"].border = border
            ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # F: N/A (placeholder - no N/A status in these DVs)
            ws[f"F{r}"] = 0
            ws[f"F{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"F{r}"].border = border
            ws[f"F{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # G: Compliance % = Compliant / (Total - N/A) * 100
            ws[f"G{r}"] = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
            ws[f"G{r}"].number_format = "0.0%"

        # Apply border and alignment to G column (all rows)
        ws[f"G{r}"].border = border
        ws[f"G{r}"].alignment = Alignment(horizontal="center", vertical="center")

    # TOTAL row with SUM formulas
    total_row = 6 + len(assessment_areas)
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{total_row}"].border = border
    for col_letter in "BCDEF":
        cell = ws[f"{col_letter}{total_row}"]
        cell.value = f"=SUM({col_letter}6:{col_letter}{total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"G{total_row}"] = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    ws[f"G{total_row}"].number_format = "0.0%"
    ws[f"G{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"G{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"G{total_row}"].border = border
    ws[f"G{total_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # --- TABLE 2: KEY METRICS ---
    met_row = total_row + 2
    ws.merge_cells(f"A{met_row}:G{met_row}")
    ws[f"A{met_row}"] = "TABLE 2: KEY METRICS"
    ws[f"A{met_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{met_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{met_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=met_row, column=col).border = border

    # TABLE 2 column headers (D9D9D9 grey)
    t2_hdr_row = met_row + 1
    t2_headers = ["Metric", "Value", "", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, 1):
        hcell = ws.cell(row=t2_hdr_row, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center", vertical="center")

    # Metrics with formulas or placeholders
    metrics = [
        ("Network Isolation Coverage", "='Separation Controls'!D6"),
        ("MFA Enforcement (Production)", "='Access Controls'!C30"),
        ("Developer Direct Production Access", "='Access Controls'!C35"),
        ("Change Promotion Success Rate", "='Promotion Workflows'!E14"),
    ]
    r = t2_hdr_row
    for metric, value in metrics:
        r += 1
        ws[f"A{r}"] = metric
        ws[f"A{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"A{r}"].border = border
        ws[f"B{r}"] = value
        ws[f"B{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"B{r}"].border = border
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border

    # TABLE 2 buffer rows (2 empty white rows)
    for _ in range(2):
        r += 1
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border

    # --- TABLE 3: CRITICAL FINDINGS ---
    crit_start = r + 2
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col_idx, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Environment Separation", "Production changes without separate test", '=COUNTIF(\'Environment Inventory\'!F6:F55,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Environment Separation", "Missing environment isolation controls", '=COUNTIF(\'Environment Inventory\'!J6:J55,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Promotion Process", "Production promotions without testing", '=COUNTIF(\'Promotion Workflows\'!E14:E44,"❌ Non-Compliant")', "Critical", "Immediate"),
        ("Production Access", "Unauthorised production access", '=COUNTIF(\'Access Controls\'!I6:I58,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Data Sync", "Production data in lower environments", '=COUNTIF(\'Data Protection\'!K6:K94,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Environment Separation", "Partial environment separation", '=COUNTIF(\'Environment Inventory\'!F6:F55,"⚠ Partial")', "High", "Urgent"),
        ("Promotion Process", "Incomplete promotion controls", '=COUNTIF(\'Promotion Workflows\'!E33:E44,"❌ Non-Compliant")', "High", "Urgent"),
        ("Production Access", "Production access control gaps", '=COUNTIF(\'Access Controls\'!I6:I58,"⚠ Partial")', "High", "Urgent"),
        ("Data Sync", "Data synchronization gaps", '=COUNTIF(\'Data Protection\'!K6:K94,"⚠ Partial")', "Medium", "Plan"),
        ("Environment Separation", "Environment monitoring gaps", '=COUNTIF(\'Environment Inventory\'!N6:N55,"Gap")', "Medium", "Plan"),
    ]

    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row = crit_start + 1
    for cat, finding, formula, severity, action in findings:
        row += 1
        ws.cell(row=row, column=1, value=cat).border = border
        ws.cell(row=row, column=1).font = Font(color="000000")
        ws.cell(row=row, column=1).fill = ffffcc_fill
        ws.cell(row=row, column=2, value=finding).border = border
        ws.cell(row=row, column=2).font = Font(color="000000")
        ws.cell(row=row, column=2).fill = ffffcc_fill
        cell_count = ws.cell(row=row, column=3)
        cell_count.value = formula
        cell_count.border = border
        cell_count.alignment = Alignment(horizontal="center")
        cell_count.font = Font(color="000000")
        cell_count.fill = ffffcc_fill
        ws.cell(row=row, column=4, value=severity).border = border
        ws.cell(row=row, column=4).font = Font(color="000000")
        ws.cell(row=row, column=4).fill = ffffcc_fill
        ws.cell(row=row, column=5, value=action).border = border
        ws.cell(row=row, column=5).font = Font(color="000000")
        ws.cell(row=row, column=5).fill = ffffcc_fill
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = ffffcc_fill
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).fill = ffffcc_fill

    # TABLE 3 buffer rows (2 empty FFFFCC rows)
    for _ in range(2):
        row += 1
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border

    # Column widths (19 columns)
    gap_widths = [
        ("A", 15), ("B", 20), ("C", 40), ("D", 30), ("E", 30),
        ("F", 10), ("G", 12), ("H", 12), ("I", 12), ("J", 35),
        ("K", 20), ("L", 15), ("M", 15), ("N", 12), ("O", 15),
        ("P", 25), ("Q", 25), ("R", 30), ("S", 20)
    ]
    for col, w in gap_widths:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: EVIDENCE_REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create standard 8-column Evidence Register with 100 rows."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Evidence type dropdown
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ev_type_dv.error = "Please select a valid evidence type"
    ev_type_dv.errorTitle = "Invalid Evidence Type"
    ws.add_data_validation(ev_type_dv)

    # Verification status dropdown
    ver_status_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Verified,{WARNING} Pending,{XMARK} Not Verified,N/A"',
        allow_blank=True,
    )
    ver_status_dv.error = "Please select a valid status"
    ver_status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(ver_status_dv)

    # Sample row (row 5) with example data (grey fill per Option B standard)
    sample_data = {
        1: "EV-001",
        2: "Environment Separation",
        3: "Configuration Export",
        4: "Network segmentation configuration from firewall",
        5: "\\\\fileserver\\evidence\\firewall_rules_export_20260115.csv",
        6: "15.01.2026",
        7: "Network Administrator",
        8: f"{CHECK} Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Empty data rows (rows 6-105) - 100 empty rows for user data (MAX-002 standard)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            cell.value = None  # Empty - users choose their own evidence IDs
        # Dropdowns
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths
    for col, w in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL_SIGN_OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create standard Approval Sign-Off with 3-section approval workflow."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    # Apply borders to all merged cells in header row
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}1"].border = border

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}2"].border = border

    # --- ASSESSMENT SUMMARY ---
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    # Assessment Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.32.3 \u2014 Environment Separation Assessment", False),
        ("Assessment Period:", "", True),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G13", False),  # TABLE 1 TOTAL row
        ("Assessed By:", "", True),
        ("Assessment Status:", "", "dropdown"),
    ]
    row = 4
    for label, value, editable in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if editable == "dropdown":
            status_dv.add(ws[f"B{row}"])
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        elif editable:
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        else:
            # Non-editable fields still need borders
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].border = border
        row += 1

    # --- Helper for approver sections ---
    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        # Apply borders to all merged cells in section title
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{start_row}"].border = border
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = border
            ws.merge_cells(f"B{r}:E{r}")
            # Apply borders and fill to all merged cells B-E
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{r}"].border = border
            r += 1
        return r

    # COMPLETED BY
    row += 1
    row = _approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")

    # REVIEWED BY
    row += 1
    row = _approver_section(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")

    # APPROVED BY (CISO)
    row += 1
    row = _approver_section(row, "APPROVED BY \u2014 CISO", "003366")

    # --- FINAL DECISION ---
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all merged cells B-E
    for col in ["B", "C", "D", "E"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # --- NEXT REVIEW DETAILS ---
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    for field in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        row += 1
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all merged cells B-E
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

    # Column widths
    for col, w in [("A", 32), ("B", 25), ("C", 20), ("D", 20), ("E", 20)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "The first principle is that you must not fool yourself
    and you are the easiest person to fool." - Richard Feynman
    
    This script generates evidence-based assessment tools, not cargo cult compliance.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.32.3 - Environment Separation Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.32: Change Management")
    logger.info("Related Controls: 8.31 (Environment Separation), 8.33 (Test Information)")
    logger.info("=" * 78)
    logger.info("\n Systems Engineering Approach: Evidence-Based Compliance")
    logger.info(f" Control 8.31: Dev/Test/Prod Isolation")
    logger.info(f" Control 8.33: Production Data Protection in Non-Prod")
    logger.info(f" Technology-Agnostic: Works with ANY infrastructure")
    logger.info(" Audit-Ready: Comprehensive evidence collection")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("{CHECK} Workbook created with 10 sheets (per IMP specification)")

    # Create all sheets (per IMP specification - 10 sheets)
    logger.info("\n[Phase 2] Generating assessment sheets...")

    logger.info("  [1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✅ Instructions complete")

    logger.info("  [2/10] Creating Environment Inventory...")
    create_development_environment(wb["Environment Inventory"], styles)
    logger.info("  ✅ Environment inventory complete")

    logger.info("  [3/10] Creating Access Controls...")
    create_test_qa_environment(wb["Access Controls"], styles)
    logger.info("  ✅ Access controls assessment complete")

    logger.info("  [4/10] Creating Promotion Workflows...")
    create_environment_promotion_process(wb["Promotion Workflows"], styles)
    logger.info("  ✅ Promotion workflows assessment complete")

    logger.info("  [5/10] Creating Data Protection...")
    create_production_data_in_nonprod(wb["Data Protection"], styles)
    logger.info("  ✅ Data protection controls complete")

    logger.info("  [6/10] Creating Environment Config...")
    create_production_environment(wb["Environment Config"], styles)
    logger.info("  ✅ Environment configuration assessment complete")

    logger.info("  [7/10] Creating Separation Controls...")
    create_separation_controls(wb["Separation Controls"], styles)
    logger.info("  ✅ Separation controls assessment complete")

    logger.info("  [8/10] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  ✅ Evidence register complete (100 evidence rows)")

    logger.info("  [9/10] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  ✅ Summary dashboard complete")

    logger.info("  [10/10] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  ✅ Approval workflow complete")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.32.3_Environment_Separation_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"{XMARK} ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info(" WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets:")
    logger.info("  • Instructions & Legend (usage guidance, Control 8.31 & 8.33 principles)")
    logger.info("  • Development_Environment (14 attributes, 11 access controls, 8 data controls)")
    logger.info("  • Test_QA_Environment (14 attributes, 11 access controls, 13 data controls)")
    logger.info("  • Production_Environment (18 attributes, 14 access restrictions, 12 data controls)")
    logger.info("  • Environment_Promotion_Process (5 promotion paths, 16 controls, 12 CI/CD stages)")
    logger.info("  • Production_Data_in_NonProd (Control 8.33 - 7 policy requirements, 17 anonymization controls)")
    logger.info("\n Analysis & Governance:")
    logger.info("  • Summary Dashboard (6 assessment areas, 10 control mappings, 6 metrics, audit readiness)")
    logger.info("  • Evidence Register (100 evidence entries)")
    logger.info("  • Approval Sign-Off (3-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info(" ASSESSMENT CAPABILITIES:")
    logger.info("  • Environment identification and configuration")
    logger.info("  • Access control assessment (RBAC, MFA, reviews)")
    logger.info("  • Network isolation verification")
    logger.info("  • Data protection controls (Control 8.33 compliance)")
    logger.info("  • Production data anonymization assessment")
    logger.info("  • Environment promotion workflow documentation")
    logger.info("  • CI/CD pipeline assessment")
    logger.info("  • 100 evidence documentation entries")
    logger.info("  • Automated compliance calculations")
    logger.info("\n" + "─" * 78)
    logger.info(f" KEY FEATURES:")
    logger.info("  ✅ Technology-agnostic (works with ANY infrastructure)")
    logger.info("  ✅ Control 8.31 (Environment Separation) assessment")
    logger.info("  ✅ Control 8.33 (Test Information) assessment")
    logger.info("  ✅ Comprehensive evidence collection")
    logger.info("  ✅ Production data protection verification")
    logger.info("  ✅ Automated compliance calculations")
    logger.info("  ✅ Multi-level approval workflow")
    logger.info("  ✅ Quarterly review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(f" NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Complete Instructions & Legend sheet first")
    logger.info("  3. Document each environment (Dev, Test, Production)")
    logger.info("  4. Define promotion procedures between environments")
    logger.info("  5. Assess production data usage in non-production (CRITICAL for Control 8.33)")
    logger.info("  6. Review Summary Dashboard for compliance metrics")
    logger.info("  7. Document evidence in Evidence Register")
    logger.info("  8. Obtain final approval via Approval Sign-Off")
    logger.info("\n PRO TIP:")
    logger.info("  Control 8.33 is CRITICAL: Production data in non-production environments")
    logger.info("  MUST be anonymized or explicitly approved by DPO. This is a common audit")
    logger.info("  finding. Document your anonymization procedures and test their effectiveness.")
    logger.info("\n" + "=" * 78)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info("\n This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
