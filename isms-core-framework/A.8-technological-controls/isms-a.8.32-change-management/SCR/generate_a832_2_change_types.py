#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.32.2 - Change Types & Categories Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Assessment Domain 2 of 4: Change Classification & Risk Management

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific change classification schemes and assessment
requirements.

Key customization areas:
1. Standard change catalog (match your pre-approved changes)
2. Change risk classification criteria (adapt to your risk appetite)
3. Emergency change triggers (customize to your business context)
4. Change calendar blackout periods (align with your business cycles)
5. Approval workflows per change type (match your governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
change classification, risk assessment, and category management against
ISO 27001:2022 Control A.8.32 requirements.

**Purpose:**
Enables systematic assessment of change type definitions, standard change
catalogs, normal change procedures, emergency change processes, risk
classification matrices, and change calendar management.

**Assessment Scope:**
- Standard change catalog and pre-authorisation criteria
- Normal change risk assessment and CAB review procedures
- Emergency change triggers and fast-track approvals
- Change risk classification (Impact × Likelihood matrices)
- Change calendar management and blackout periods
- Change type compliance verification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and classification standards
2. Standard Changes Catalog - Pre-approved low-risk changes
3. Normal_Changes_Assessment - CAB-reviewed changes and procedures
4. Emergency_Changes - Fast-track emergency change process
5. Change_Risk_Classification - Risk matrix and scoring
6. Change Calendar Management - Blackout periods and scheduling
7. Summary Dashboard - Compliance metrics and analytics
8. Evidence Register - Audit evidence tracking
9. Approval Sign-Off - Stakeholder approval workflow

**Key Features:**
- Technology-agnostic assessment (works with any ITSM tool)
- Standard change catalog template with customization guidance
- Risk matrix with Impact × Likelihood methodology
- Change calendar integration for blackout period management
- Automated compliance calculations
- Evidence linkage for audit traceability

**Integration:**
consolidates data from all four change management assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a832_2_change_types.py

Output:
    File: ISMS_A_8_32_2_Change_Types_Categories_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Define your organisation's standard change catalog
    2. Document normal change risk assessment procedures
    3. Define emergency change triggers specific to your business
    4. Customize risk matrix to your risk appetite
    5. Document change calendar blackout periods
    6. Review Summary Dashboard for compliance metrics
    7. Collect and link audit evidence

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32
Assessment Domain:    2 of 4 (Change Classification & Risk Management)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.2: Change Types & Categories Implementation Guide
    - ISMS-IMP-A.8.32.1: Change Process Assessment (Domain 1)
    - ISMS-IMP-A.8.32.3: Environment Separation Assessment (Domain 3)
    - ISMS-IMP-A.8.32.4: Testing & Validation Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.32.2 specification
    - Supports comprehensive change classification evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Standard Change Philosophy:**
Standard changes are pre-approved, low-risk, well-understood changes that follow
a documented procedure. They don't require CAB approval but DO require logging
and monitoring for compliance.

**Risk Classification Flexibility:**
Customize the risk matrix to your organisation's risk appetite. A startup might
accept higher risks than a regulated financial institution. The framework is
intentionally flexible.

**Emergency Change Reality:**
Emergency changes will happen. The goal is controlled urgency, not prevention.
Document clear triggers, fast-track approvals, and mandatory post-implementation
reviews.

**Audit Considerations:**
Auditors will verify that change types are clearly defined, risk assessments
are systematic, and emergency changes have proper justification and review.

**Data Protection:**
Assessment workbooks contain sensitive information about change management
practices, risk tolerances, and operational procedures. Handle according to
your data classification policies.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.32.2"
WORKBOOK_NAME = "Change Types & Categories Assessment"
CONTROL_ID = "A.8.32"
CONTROL_NAME = "Change Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠  Warning sign
GEAR = '\u2699'       # ⚙  Gear
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP-A.8.32.2 spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.32.2 specification (10 sheets)
    sheets = [
        "Instructions & Legend",
        "Standard Changes Catalog",
        "Normal Change Classification",
        "Emergency Change Procedures",
        "Risk Assessment Matrix",
        "Change Calendar Management",
        "Classification Metrics",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "risk_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "risk_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "risk_high": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "risk_critical": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell. Creates NEW style objects."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    CUSTOMIZE: Adapt dropdown values for change classification context.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'definition_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Defined,⚠ Partial,❌ Not Defined, Planned,N/A"',
            allow_blank=False
        ),
        'change_category': DataValidation(
            type="list",
            formula1='"Infrastructure,Application,Security,Data,Network,Cloud,User Access,Configuration,Other"',
            allow_blank=False
        ),
        'frequency': DataValidation(
            type="list",
            formula1='"Daily,Weekly,Monthly,Quarterly,Annual,On-Demand,Rare"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Low,Medium,High,Critical"',
            allow_blank=False
        ),
        'std_change_risk': DataValidation(
            type="list",
            formula1='"Low,Medium"',
            allow_blank=False
        ),
        'std_change_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Active,⚠ Under Review,❌ Retired, Proposed"',
            allow_blank=False
        ),
        'cab_requirement': DataValidation(
            type="list",
            formula1='"Mandatory,Recommended,Optional,Not Required"',
            allow_blank=False
        ),
        'approval_authority': DataValidation(
            type="list",
            formula1='"Change Manager,CAB,Service Owner,CISO,CIO,Multiple Required,Other"',
            allow_blank=False
        ),
        'implementation_level': DataValidation(
            type="list",
            formula1=f'"{CHECK} Always,⚠ Sometimes,❌ Never, Planned"',
            allow_blank=False
        ),
        'exceptions_allowed': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,❌ No,⚠ Case-by-Case,N/A"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,⚠ Partial,❌ No"',
            allow_blank=False
        ),
        'emergency_criteria_definition': DataValidation(
            type="list",
            formula1=f'"{CHECK} Clearly Defined,⚠ Partially,❌ Not Defined"',
            allow_blank=False
        ),
        'user_count': DataValidation(
            type="list",
            formula1='"<10 users,10-50,50-100,100-1000,>1000 users,All users,Business-critical,Other"',
            allow_blank=False
        ),
        'downtime_potential': DataValidation(
            type="list",
            formula1='"<15 min,15-60 min,1-2 hours,2-8 hours,>8 hours,Irreversible,Business-critical,Other"',
            allow_blank=False
        ),
        'data_loss_risk': DataValidation(
            type="list",
            formula1='"None,Minimal,Moderate,High"',
            allow_blank=False
        ),
        'complexity': DataValidation(
            type="list",
            formula1='"Simple,Moderate,Complex,Very Complex"',
            allow_blank=False
        ),
        'testing_maturity': DataValidation(
            type="list",
            formula1='"Extensively Tested,Well Tested,Limited Testing,Untested"',
            allow_blank=False
        ),
        'convening_method': DataValidation(
            type="list",
            formula1='"Conference Call,Video,IM Group,Emergency Hotline,Other"',
            allow_blank=False
        ),
        'applicable_changes': DataValidation(
            type="list",
            formula1='"All,Standard Only,Normal Only,Emergency Only,Custom"',
            allow_blank=False
        ),
        'blackout_reason': DataValidation(
            type="list",
            formula1='"Financial Close,Holiday Period,Peak Business,Audit,Other"',
            allow_blank=False
        ),
        'conflict_detection': DataValidation(
            type="list",
            formula1='"Automated,Manual,None"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Change Request,Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other,N/A"',
            allow_blank=True
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⚠ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Final,⚠ Requires Remediation, Draft,❌ Re-assessment Required"',
            allow_blank=False
        ),
        'review_recommendation': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approve,⚠ Approve with Conditions,❌ Reject, Require Rework"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approved,⚠ Approved with Conditions,❌ Rejected"',
            allow_blank=False
        ),
        'regulatory_review': DataValidation(
            type="list",
            formula1='"Yes,No,To Be Determined"',
            allow_blank=False
        ),
    }

    # NOTE: Do NOT add validations here. They are added per-sheet via
    # finalize_validations() AFTER cells are assigned, to avoid empty
    # sqref entries that trigger Excel "Repaired" warnings.

    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete the Standard Changes Catalog with YOUR pre-approved changes.', '2. Document YOUR normal change assessment criteria.', '3. Define YOUR emergency change triggers and procedures.', '4. Configure YOUR change risk classification matrix.', '5. Document YOUR change calendar management approach.', '6. Review the Summary Dashboard for compliance metrics.', '7. Maintain the Evidence Register for audit traceability.', '8. Obtain final approval and sign-off.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_standard_changes_catalog(ws, styles):
    """Create Standard Changes Catalog sheet with 50 pre-formatted entries."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:M1")
    ws["A1"] = "STANDARD CHANGES CATALOG"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:M2")
    ws["A2"] = "Pre-approved, low-risk changes with documented procedures"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Change ID", "Change Title", "Description", "Category", "Frequency", "Pre-requisites", 
               "Procedure Location", "Owner", "Approval Date", "Review Date", "Risk Level", "Status", "Evidence"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Sample row (row 5) with example data
    row += 1
    sample_data = {
        1: "STD-001",
        2: "Restart Application Server",
        3: "Standard procedure to restart application server after configuration changes",
        4: "Application",
        5: "Weekly",
        6: "Business approval, maintenance window scheduled",
        7: "\\\\procedures\\app_server_restart.pdf",
        8: "Application Support Team",
        9: "15.01.2025",
        10: "15.01.2026",
        11: "Low",
        12: "Active",
        13: "EV-APP-001",
    }
    for col_idx, value in sample_data.items():
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = styles["border"]
        if col_idx in [9, 10]:  # Date columns
            cell.number_format = 'DD.MM.YYYY'

    # Add dropdowns to sample row
    validations['change_category'].add(ws[f"D{row}"])
    validations['frequency'].add(ws[f"E{row}"])
    validations['std_change_risk'].add(ws[f"K{row}"])
    validations['std_change_status'].add(ws[f"L{row}"])
    validations['evidence_type'].add(ws.cell(row=row, column=13))

    # Empty data rows (rows 6-55) - 50 rows for user data
    row += 1
    for i in range(50):
        # All columns: FFFFCC fill + border, NO pre-filled values
        for col_idx in range(1, 14):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]
            cell.value = None  # Empty - users define their own standard changes
            if col_idx in [9, 10]:  # Date columns
                cell.number_format = 'DD.MM.YYYY'

        # Add dropdowns to empty rows
        validations['change_category'].add(ws[f"D{row}"])
        validations['frequency'].add(ws[f"E{row}"])
        validations['std_change_risk'].add(ws[f"K{row}"])
        validations['std_change_status'].add(ws[f"L{row}"])
        validations['evidence_type'].add(ws.cell(row=row, column=13))

        row += 1

    # Summary Metrics
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "STANDARD CHANGE SUMMARY METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Value", "Notes"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        "Total Standard Changes Defined",
        "Active Standard Changes",
        "Standard Changes Under Review",
        "Standard Changes Requiring Annual Review",
        "Most Common Category",
        "Average Age of Standard Changes",
    ]

    row += 1
    for metric in metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"B{row}"].value = "[FORMULA]"
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 25

    ws.freeze_panes = "A4"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: NORMAL CHANGES ASSESSMENT SHEET
# ============================================================================

def create_normal_changes_assessment(ws, styles):
    """Create Normal_Changes_Assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "NORMAL CHANGES ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Changes requiring risk assessment, approval, and CAB review"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== NORMAL CHANGE CRITERIA ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "NORMAL CHANGE CRITERIA"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    criteria_headers = ["Criterion", "Defined", "Documentation Reference", "Assessment Method", "Responsible Role", "Compliance", "Evidence"]
    for col_idx, header in enumerate(criteria_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    criteria = [
        "Change does not meet Standard criteria",
        "Change is not an emergency",
        "Risk assessment required",
        "Impact assessment required",
        "CAB review required (based on risk)",
        "Business justification required",
        "Implementation plan required",
        "Test plan required",
        "Rollback plan required",
        "Communication plan required",
        "PIR (Post-Implementation Review) required",
        "Change window scheduling required",
    ]

    row += 1
    for criterion in criteria:
        ws[f"A{row}"] = criterion

        # Editable cells
        for col in ["C", "D", "E", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Dropdowns
        validations['definition_status'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['compliance_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        # Evidence dropdown (col 7 = G)
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # ==================== RISK-BASED APPROVAL PATHS ====================
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "RISK-BASED APPROVAL PATHS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    approval_headers = ["Risk Category", "Impact Category", "Approval Authority", "CAB Review", "Typical Timeline", "Success Rate", "Documented", "Evidence"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    risk_combinations = [
        ("Low", "Low"), ("Low", "Medium"), ("Low", "High"),
        ("Medium", "Low"), ("Medium", "Medium"), ("Medium", "High"),
        ("High", "Low"), ("High", "Medium"), ("High", "High"),
        ("Critical", "Any"),
    ]

    row += 1
    for risk, impact in risk_combinations:
        ws[f"A{row}"] = risk
        ws[f"B{row}"] = impact

        # Editable cells
        for col in ["E", "F", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Dropdowns
        validations['approval_authority'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['cab_requirement'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        validations['definition_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        # Evidence dropdown (col 8 = H)
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: EMERGENCY CHANGES SHEET
# ============================================================================

def create_emergency_changes(ws, styles):
    """Create Emergency_Changes sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "EMERGENCY CHANGES"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Urgent changes meeting specific emergency criteria with E-CAB approval"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== EMERGENCY CRITERIA DEFINITION ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "EMERGENCY CRITERIA DEFINITION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    criteria_headers = ["Emergency Criterion", "Defined", "Specific Triggers", "Escalation Path", "Response Time SLA", "Documented", "Evidence"]
    # Adjust to 7 columns
    for col_idx, header in enumerate(criteria_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    emergency_criteria = [
        "System Outage (Critical Services)",
        "Security Incident Response",
        "Data Loss Prevention",
        "Regulatory Compliance Deadline",
        "Health & Safety Risk",
        "Business Continuity Threat",
        "[Custom Criterion 1]",
        "[Custom Criterion 2]",
    ]

    row += 1
    for criterion in emergency_criteria:
        ws[f"A{row}"] = criterion

        # Editable cells
        for col in ["C", "D", "E", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Dropdowns
        validations['emergency_criteria_definition'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['definition_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        # Evidence dropdown (col 7 = G)
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # ==================== EMERGENCY CHANGE PROCESS REQUIREMENTS ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGE PROCESS REQUIREMENTS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    req_headers = ["Requirement", "Implemented", "Process Description", "Responsible Role", "Exceptions Allowed", "Compliance", "Evidence"]
    for col_idx, header in enumerate(req_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    requirements = [
        "Emergency criteria must be met",
        "E-CAB convened",
        "Minimum E-CAB members defined",
        "Emergency approval documented",
        "Risk acceptance explicit",
        "Communication immediate",
        "Implementation window immediate",
        "Rollback plan prepared (where feasible)",
        "Incident ticket linked",
        "Mandatory PIR within 2 business days",
        "Retrospective CAB review",
        "Emergency change metrics tracked",
    ]

    row += 1
    for requirement in requirements:
        ws[f"A{row}"] = requirement

        # Editable cells
        for col in ["C", "D", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Dropdowns
        validations['implementation_level'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['exceptions_allowed'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        validations['compliance_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        # Evidence dropdown (col 7 = G)
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # ==================== EMERGENCY CHANGE METRICS ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGE METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Target", "Current (Last Quarter)", "Status", "Notes"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        ("Emergency changes as % of total changes",  "<5%"),
        ("Average E-CAB response time",              "<4 hours"),
        ("Emergency changes with PIR completed",     "100%"),
        ("Emergency changes leading to incidents",   "0%"),
        ("Retrospective CAB review completion",      "100%"),
        ("False emergency declarations",             "<5%"),
        ("Emergency change success rate",            ">95%"),
    ]

    row += 1
    for metric_name, target in metrics:
        ws[f"A{row}"] = metric_name
        ws[f"B{row}"] = target

        # Target already written above; Notes - editable
        for col in ["B", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Current - editable
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Status - editable (manual assessment; targets are text-based)
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 7: CHANGE RISK CLASSIFICATION SHEET
# ============================================================================

def create_change_risk_classification(ws, styles):
    """Create Change_Risk_Classification sheet with risk matrix."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE RISK CLASSIFICATION MATRIX"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Risk = Impact × Likelihood methodology"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== IMPACT LEVEL DEFINITIONS ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "IMPACT LEVEL DEFINITIONS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    impact_headers = ["Impact Level", "User Count Affected", "Service Downtime Potential", "Recovery Time", "Data Loss Risk", "Financial Impact", "Documented", "Evidence"]
    for col_idx, header in enumerate(impact_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    impact_levels = ["Low", "Medium", "High", "Critical"]

    row += 1
    for level in impact_levels:
        ws[f"A{row}"] = level

        # Editable cells
        for col in ["D", "F", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Dropdowns
        validations['user_count'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['downtime_potential'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['data_loss_risk'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        validations['definition_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        # Evidence dropdown (col 8 = H)
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # ==================== LIKELIHOOD LEVEL DEFINITIONS ====================
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "LIKELIHOOD LEVEL DEFINITIONS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    likelihood_headers = ["Likelihood Level", "Failure Probability", "Complexity", "Dependencies", "Testing Maturity", "Historical Success Rate", "Documented", "Evidence"]
    for col_idx, header in enumerate(likelihood_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    likelihood_levels = ["Low", "Medium", "High"]

    row += 1
    for level in likelihood_levels:
        ws[f"A{row}"] = level

        # Editable cells
        for col in ["B", "D", "F", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Dropdowns
        validations['complexity'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['testing_maturity'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        validations['definition_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        # Evidence dropdown (col 8 = H)
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # ==================== RISK MATRIX ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "RISK MATRIX (Impact × Likelihood)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    # Risk matrix headers
    ws[f"A{row}"] = ""
    ws[f"B{row}"] = "Low Impact"
    ws[f"C{row}"] = "Medium Impact"
    ws[f"D{row}"] = "High Impact"
    ws[f"E{row}"] = "Critical Impact"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Risk matrix data with color coding
    risk_matrix = [
        ("Low Likelihood", "LOW RISK", "LOW RISK", "MEDIUM RISK", "HIGH RISK"),
        ("Medium Likelihood", "LOW RISK", "MEDIUM RISK", "HIGH RISK", "CRITICAL RISK"),
        ("High Likelihood", "MEDIUM RISK", "HIGH RISK", "CRITICAL RISK", "CRITICAL RISK"),
    ]

    row += 1
    for likelihood, low, medium, high, critical in risk_matrix:
        ws[f"A{row}"] = likelihood
        ws[f"A{row}"].font = Font(bold=True)
        
        ws[f"B{row}"] = low
        ws[f"C{row}"] = medium
        ws[f"D{row}"] = high
        ws[f"E{row}"] = critical
        
        # Apply risk color coding
        for col, value in [("B", low), ("C", medium), ("D", high), ("E", critical)]:
            if "LOW" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif "MEDIUM" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif "HIGH" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif "CRITICAL" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col}{row}"].font = Font(bold=True)
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: CHANGE CALENDAR MANAGEMENT SHEET
# ============================================================================

def create_change_calendar_management(ws, styles):
    """Create Change Calendar Management sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE CALENDAR MANAGEMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Scheduling, blackout windows, and conflict detection"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== CHANGE WINDOW DEFINITIONS ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "CHANGE WINDOW DEFINITIONS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    window_headers = ["Change Window", "Days/Times", "Applicable Change Types", "Approval Required", "Advance Notice", "Documented", "Evidence"]
    for col_idx, header in enumerate(window_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    change_windows = [
        "Standard Maintenance Window",
        "Business Hours (Restricted)",
        "After-Hours Window",
        "Emergency Window",
        "[Custom Window 1]",
        "[Custom Window 2]",
    ]

    row += 1
    for window in change_windows:
        ws[f"A{row}"] = window

        # Editable cells
        for col in ["B", "E", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = styles["border"]

        # Dropdowns
        validations['applicable_changes'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        validations['definition_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        # Evidence dropdown (col 7 = G)
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # ==================== BLACKOUT PERIODS ====================
    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "BLACKOUT PERIODS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    blackout_headers = ["Blackout Period", "Start Date", "End Date", "Reason", "Affected Systems/Services", "Exceptions Allowed", "Exception Approver", "Documented", "Evidence"]
    # Adjust to 9 columns
    for col_idx, header in enumerate(blackout_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Sample row with example data
    row += 1
    sample_data = {
        1: "Year-End Freeze",
        2: "20.12.2025",
        3: "05.01.2026",
        4: "Financial Close",
        5: "All production systems",
        6: "Critical only",
        7: "CTO",
        8: "Documented",
        9: "EV-BL-001",
    }
    for col_idx, value in sample_data.items():
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = styles["border"]
        if col_idx in [2, 3]:  # Date columns
            cell.number_format = 'DD.MM.YYYY'

    # Add dropdowns to sample row
    validations['blackout_reason'].add(ws[f"D{row}"])
    validations['exceptions_allowed'].add(ws[f"F{row}"])
    validations['definition_status'].add(ws[f"H{row}"])
    validations['evidence_type'].add(ws.cell(row=row, column=9))

    # Empty data rows (50 rows for user data)
    row += 1
    for i in range(50):
        for col_idx in range(1, 10):  # Columns A-I
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]
            cell.value = None  # Empty - users define their own blackout periods
            if col_idx in [2, 3]:  # Date columns
                cell.number_format = 'DD.MM.YYYY'

        # Add dropdowns to empty rows
        validations['blackout_reason'].add(ws[f"D{row}"])
        validations['exceptions_allowed'].add(ws[f"F{row}"])
        validations['definition_status'].add(ws[f"H{row}"])
        validations['evidence_type'].add(ws.cell(row=row, column=9))

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8.5: CLASSIFICATION METRICS SHEET
# ============================================================================

def create_classification_metrics(ws, styles):
    """Create Classification Metrics sheet tracking change classification metrics."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE CLASSIFICATION METRICS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Track metrics related to change type distribution and classification accuracy"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== CLASSIFICATION METRICS ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CHANGE TYPE DISTRIBUTION METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    headers = ["Metric", "Target", "Current", "Period", "Trend", "Status", "Owner", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    metrics = [
        ("Standard Changes %", ">60%", "", "Monthly", "", "", "", ""),
        ("Normal Changes %", "25-35%", "", "Monthly", "", "", "", ""),
        ("Emergency Changes %", "<5%", "", "Monthly", "", "", "", ""),
        ("Classification Accuracy", ">95%", "", "Monthly", "", "", "", ""),
        ("Re-Classification Rate", "<2%", "", "Monthly", "", "", "", ""),
        ("Standard Catalog Utilization", ">80%", "", "Monthly", "", "", "", ""),
    ]

    for metric_data in metrics:
        for col_idx, value in enumerate(metric_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles["input_cell"])
        row += 1

    row += 2

    # ==================== CLASSIFICATION ACCURACY ====================
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "CLASSIFICATION ACCURACY TRACKING"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    accuracy_headers = ["Period", "Total Changes", "Correctly Classified", "Misclassified", "Accuracy %", "Review Date", "Reviewer"]
    for col_idx, header in enumerate(accuracy_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    # Sample row with example data
    sample_data = {
        1: "2025 Q4",
        2: "142",
        3: "139",
        4: "3",
        5: "97.9%",
        6: "15.01.2026",
        7: "Change Manager",
    }
    for col_idx, value in sample_data.items():
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["input_cell"])  # FFFFCC — standard yellow sample row
    row += 1

    # Empty data rows (50 rows for user data)
    for _ in range(50):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, styles["input_cell"])
            cell.value = None  # Empty - users track their own periods
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with standard compliance table and metrics."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE TYPES ASSESSMENT — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.32: Change Management"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # --- TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW ---
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    assessment_areas = [
        "Standard Changes Catalog",
        "Normal Change Classification",
        "Emergency Change Procedures",
        "Risk Assessment Matrix - Impact Levels",
        "Risk Assessment Matrix - Likelihood Levels",
        "Change Calendar Management - Change Windows",
        "Change Calendar Management - Blackout Periods",
    ]

    # Area configurations: (sheet_name, status_col, [compliant, partial, non_compliant], row_start, row_end)
    # None = manual entry area
    area_configs = [
        ('Standard Changes Catalog', 'L', ['✅ Active', '⚠ Under Review', '❌ Retired'], 6, 65),  # Row 5 is sample (grey), data 6-65
        ('Normal Change Classification', 'F', ['✅ Yes', '⚠ Partial', '❌ No'], 6, 55),  # Pre-filled, no sample
        ('Emergency Change Procedures', 'B', ['✅ Always', '⚠ Sometimes', '❌ Never'], 17, 40),  # Process Requirements section (extends to row 40)
        # Risk Assessment Matrix - 2 sections with formulas (extended to cover all data to row 16)
        ('Risk Assessment Matrix', 'G', ['✅ Defined', '⚠ Partial', '❌ Not Defined'], 6, 16),  # Impact Levels section
        ('Risk Assessment Matrix', 'G', ['✅ Defined', '⚠ Partial', '❌ Not Defined'], 14, 16),  # Likelihood Levels section
        # Change Calendar Management - 2 sections with formulas (extended to cover all data to row 66)
        ('Change Calendar Management', 'F', ['✅ Defined', '⚠ Partial', '❌ Not Defined'], 6, 66),  # Change Windows section
        ('Change Calendar Management', 'H', ['✅ Defined', '⚠ Partial', '❌ Not Defined'], 16, 66),  # Blackout Periods section (sample at row 16)
    ]

    for i, area_name in enumerate(assessment_areas):
        r = 6 + i
        ws[f"A{r}"] = area_name
        ws[f"A{r}"].font = Font(name="Calibri", size=10)
        ws[f"A{r}"].border = border

        if area_configs[i] is None:
            # MANUAL ENTRY - use placeholder cells
            for col in "BCDEF":
                cell = ws[f"{col}{r}"]
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Manual areas get placeholder in column G
            ws[f"G{r}"] = "[enter %]"
            ws[f"G{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        else:
            # FORMULA-BASED - generate compliance formulas
            sheet_name, status_col, status_values, row_start, row_end = area_configs[i]
            compliant_val, partial_val, non_compliant_val = status_values

            # B: Total items (count non-empty rows in column A)
            ws[f"B{r}"] = f"=COUNTA('{sheet_name}'!A{row_start}:A{row_end})"
            ws[f"B{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"B{r}"].border = border
            ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # C: Compliant
            ws[f"C{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{compliant_val}\")"
            ws[f"C{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"C{r}"].border = border
            ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # D: Partially Compliant
            ws[f"D{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{partial_val}\")"
            ws[f"D{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"D{r}"].border = border
            ws[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # E: Non-Compliant (including "Proposed"/"Planned" for specific sheets)
            if area_name == "Standard Changes Catalog":
                # Include "Proposed" as non-compliant
                ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")+COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"Proposed\")"
            elif area_name == "Emergency Change Procedures":
                # Include "Planned" as non-compliant (ISO requires implementation, not planning)
                ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")+COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"Planned\")"
            elif "Risk Assessment Matrix" in area_name or "Change Calendar Management" in area_name:
                # Include "Planned" as non-compliant for definition_status DV
                ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")+COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"Planned\")"
            else:
                ws[f"E{r}"] = f"=COUNTIF('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end},\"{non_compliant_val}\")"
            ws[f"E{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"E{r}"].border = border
            ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # F: N/A (placeholder - no N/A status in these DVs)
            ws[f"F{r}"] = 0
            ws[f"F{r}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f"F{r}"].border = border
            ws[f"F{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # G: Compliance % = Compliant / (Total - N/A) * 100
            ws[f"G{r}"] = f'=IF((B{r}-F{r})=0,0,C{r}/(B{r}-F{r}))'
            ws[f"G{r}"].number_format = "0.0%"

        # Apply border and alignment to G column (all rows)
        ws[f"G{r}"].border = border
        ws[f"G{r}"].alignment = Alignment(horizontal="center", vertical="center")

    # TOTAL row with SUM formulas
    total_row = 6 + len(assessment_areas)
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{total_row}"].border = border
    for col_letter in "BCDEF":
        cell = ws[f"{col_letter}{total_row}"]
        cell.value = f"=SUM({col_letter}6:{col_letter}{total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"G{total_row}"] = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    ws[f"G{total_row}"].number_format = "0.0%"
    ws[f"G{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"G{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"G{total_row}"].border = border
    ws[f"G{total_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # --- TABLE 2: KEY METRICS ---
    met_row = total_row + 2
    ws.merge_cells(f"A{met_row}:G{met_row}")
    ws[f"A{met_row}"] = "TABLE 2: KEY METRICS"
    ws[f"A{met_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{met_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{met_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=met_row, column=col).border = border

    # TABLE 2 column headers (D9D9D9 grey)
    t2_hdr_row = met_row + 1
    t2_headers = ["Metric", "Value", "", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, 1):
        hcell = ws.cell(row=t2_hdr_row, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center", vertical="center")

    # Metrics with formulas or placeholders
    metrics = [
        ("Standard Change Utilisation Rate", "='Classification Metrics'!C6"),
        ("Emergency Change Rate", "='Classification Metrics'!C8"),
        ("Classification Accuracy", "='Classification Metrics'!C9"),
        ("Risk Matrix Coverage", "[enter %]"),
    ]
    r = t2_hdr_row
    for metric, value in metrics:
        r += 1
        ws[f"A{r}"] = metric
        ws[f"A{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"A{r}"].border = border
        ws[f"B{r}"] = value
        ws[f"B{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"B{r}"].border = border
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border

    # TABLE 2 buffer rows (2 empty white rows)
    for _ in range(2):
        r += 1
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border

    # --- TABLE 3: CRITICAL FINDINGS ---
    crit_start = r + 2
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col_idx, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Standard Changes", "Standard changes without pre-authorisation", '=COUNTIF(\'Standard Changes Catalog\'!L6:L65,"❌ Retired")', "Critical", "Immediate"),
        ("Normal Changes", "Normal changes without CAB approval", '=COUNTIF(\'Normal Change Classification\'!F6:F55,"❌ No")', "Critical", "Immediate"),
        ("Emergency Changes", "Emergency changes without justification", '=COUNTIF(\'Emergency Change Procedures\'!I6:I55,"❌ Not Implemented")', "Critical", "Immediate"),
        ("Emergency Changes", "Emergency changes exceeding threshold", '=COUNTIF(\'Emergency Change Procedures\'!O6:O55,"Above Threshold")', "Critical", "Immediate"),
        ("Normal Changes", "High-risk changes without risk assessment", '=COUNTIF(\'Normal Change Classification\'!J6:J55,"❌ No")', "Critical", "Immediate"),
        ("Standard Changes", "Standard changes with partial authorisation", '=COUNTIF(\'Standard Changes Catalog\'!F6:F65,"⚠ Partial")', "High", "Urgent"),
        ("Normal Changes", "Normal changes with incomplete approval", '=COUNTIF(\'Normal Change Classification\'!F6:F55,"⚠ Partial")', "High", "Urgent"),
        ("Emergency Changes", "Emergency changes with partial retrospective approval", '=COUNTIF(\'Emergency Change Procedures\'!I6:I55,"⚠ Partial")', "High", "Urgent"),
        ("Standard Changes", "Standard changes requiring review", '=COUNTIF(\'Standard Changes Catalog\'!O6:O65,"Review Required")', "Medium", "Plan"),
        ("Normal Changes", "Normal changes with minor gaps", '=COUNTIF(\'Normal Change Classification\'!O6:O55,"Minor Gap")', "Medium", "Plan"),
    ]

    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row = crit_start + 1
    for cat, finding, formula, severity, action in findings:
        row += 1
        ws.cell(row=row, column=1, value=cat).border = border
        ws.cell(row=row, column=1).font = Font(color="000000")
        ws.cell(row=row, column=1).fill = ffffcc_fill
        ws.cell(row=row, column=2, value=finding).border = border
        ws.cell(row=row, column=2).font = Font(color="000000")
        ws.cell(row=row, column=2).fill = ffffcc_fill
        cell_count = ws.cell(row=row, column=3)
        cell_count.value = formula
        cell_count.border = border
        cell_count.alignment = Alignment(horizontal="center")
        cell_count.font = Font(color="000000")
        cell_count.fill = ffffcc_fill
        ws.cell(row=row, column=4, value=severity).border = border
        ws.cell(row=row, column=4).font = Font(color="000000")
        ws.cell(row=row, column=4).fill = ffffcc_fill
        ws.cell(row=row, column=5, value=action).border = border
        ws.cell(row=row, column=5).font = Font(color="000000")
        ws.cell(row=row, column=5).fill = ffffcc_fill
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = ffffcc_fill
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).fill = ffffcc_fill

    # TABLE 3 buffer rows (2 empty FFFFCC rows)
    for _ in range(2):
        row += 1
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border

    # Column widths (19 columns)
    gap_widths = [
        ("A", 15), ("B", 20), ("C", 40), ("D", 30), ("E", 30),
        ("F", 10), ("G", 12), ("H", 12), ("I", 12), ("J", 35),
        ("K", 20), ("L", 15), ("M", 15), ("N", 12), ("O", 15),
        ("P", 25), ("Q", 25), ("R", 30), ("S", 20)
    ]
    for col, w in gap_widths:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create standard 8-column Evidence Register with 100 rows."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Evidence type dropdown
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ev_type_dv.error = "Please select a valid evidence type"
    ev_type_dv.errorTitle = "Invalid Evidence Type"
    ws.add_data_validation(ev_type_dv)

    # Verification status dropdown
    ver_status_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Verified,{WARNING} Pending,{XMARK} Not Verified,N/A"',
        allow_blank=True,
    )
    ver_status_dv.error = "Please select a valid status"
    ver_status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(ver_status_dv)

    # Sample row (row 5) with example data (grey fill per Option B standard)
    sample_data = {
        1: "EV-001",
        2: "Change Types Assessment",
        3: "System Screenshot",
        4: "Standard change catalog export from ITSM tool",
        5: "\\\\fileserver\\evidence\\std_change_catalog_20260115.xlsx",
        6: "15.01.2026",
        7: "Change Administrator",
        8: f"{CHECK} Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Empty data rows (rows 6-105) - 100 empty rows for user data (MAX-002 standard)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            cell.value = None  # Empty - users choose their own evidence IDs
        # Dropdowns
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths
    for col, w in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create standard Approval Sign-Off with 3-section approval workflow."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    # Apply borders to all merged cells in header row
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}1"].border = border

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}2"].border = border

    # --- ASSESSMENT SUMMARY ---
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    # Assessment Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.32.2 \u2014 Change Types & Categories Assessment", False),
        ("Assessment Period:", "", True),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G13", False),  # TABLE 1 TOTAL row
        ("Assessed By:", "", True),
        ("Assessment Status:", "", "dropdown"),
    ]
    row = 4
    for label, value, editable in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if editable == "dropdown":
            status_dv.add(ws[f"B{row}"])
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        elif editable:
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        else:
            # Non-editable fields still need borders
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].border = border
        row += 1

    # --- Helper for approver sections ---
    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        # Apply borders to all merged cells in section title
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{start_row}"].border = border
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = border
            ws.merge_cells(f"B{r}:E{r}")
            # Apply borders and fill to all merged cells B-E
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{r}"].border = border
            r += 1
        return r

    # COMPLETED BY
    row += 1
    row = _approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")

    # REVIEWED BY
    row += 1
    row = _approver_section(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")

    # APPROVED BY (CISO)
    row += 1
    row = _approver_section(row, "APPROVED BY \u2014 CISO", "003366")

    # --- FINAL DECISION ---
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all merged cells B-E
    for col in ["B", "C", "D", "E"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # --- NEXT REVIEW DETAILS ---
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    for field in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        row += 1
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all merged cells B-E
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

    # Column widths
    for col, w in [("A", 32), ("B", 25), ("C", 20), ("D", 20), ("E", 20)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: Create evidence-based assessment tools for change classification,
    not checkbox compliance theater.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.32.2 - Change Types & Categories Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.32: Change Management")
    logger.info("=" * 78)
    logger.info("\n Systems Engineering Approach: Evidence-Based Compliance")
    logger.info(f" Technology-Agnostic: Works with ANY change classification approach")
    logger.info(f" Audit-Ready: Comprehensive evidence collection")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("{CHECK} Workbook created with 9 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")

    logger.info("  [1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✅ Instructions complete")

    logger.info("  [2/10] Creating Standard Changes Catalog...")
    create_standard_changes_catalog(wb["Standard Changes Catalog"], styles)
    logger.info("  ✅ Standard changes catalog complete (50 entries)")

    logger.info("  [3/10] Creating Normal Change Classification...")
    create_normal_changes_assessment(wb["Normal Change Classification"], styles)
    logger.info("  ✅ Normal change classification complete")

    logger.info("  [4/10] Creating Emergency Change Procedures...")
    create_emergency_changes(wb["Emergency Change Procedures"], styles)
    logger.info("  ✅ Emergency change procedures complete")

    logger.info("  [5/10] Creating Risk Assessment Matrix...")
    create_change_risk_classification(wb["Risk Assessment Matrix"], styles)
    logger.info("  ✅ Risk assessment matrix complete")

    logger.info("  [6/10] Creating Change Calendar Management...")
    create_change_calendar_management(wb["Change Calendar Management"], styles)
    logger.info("  ✅ Change calendar and blackout periods complete")

    logger.info("  [7/10] Creating Classification Metrics...")
    create_classification_metrics(wb["Classification Metrics"], styles)
    logger.info("  ✅ Classification metrics complete")

    logger.info("  [8/10] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  ✅ Evidence register complete (100 evidence rows)")

    logger.info("  [9/10] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  ✅ Dashboard complete (compliance metrics)")

    logger.info("  [10/10] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  ✅ Approval workflow complete (3-level sign-off)")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.32.2_Change_Types_Categories_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"{XMARK} ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info(" WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets (10 per IMP specification):")
    logger.info("  • Instructions & Legend (usage guidance)")
    logger.info("  • Standard Changes Catalog (50 pre-approved changes)")
    logger.info("  • Normal Change Classification (risk-based approval paths)")
    logger.info("  • Emergency Change Procedures (E-CAB procedures, <5% target)")
    logger.info("  • Risk Assessment Matrix (Impact x Likelihood matrix)")
    logger.info("  • Change Calendar Management (blackout periods)")
    logger.info("  • Classification Metrics (distribution and accuracy)")
    logger.info("\n Analysis & Governance:")
    logger.info("  • Evidence Register (100 evidence entries)")
    logger.info("  • Summary Dashboard (compliance metrics)")
    logger.info("  • Approval Sign-Off (3-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info(" ASSESSMENT CAPABILITIES:")
    logger.info("  • 50 standard changes catalog entries")
    logger.info("  • Risk matrix (4 impact × 3 likelihood levels)")
    logger.info("  • Emergency change metrics (<5% target)")
    logger.info("  • Change calendar with blackout periods")
    logger.info("  • 12 policy requirements mapped")
    logger.info("  • 100 evidence documentation entries")
    logger.info("  • Automated compliance calculations")
    logger.info("\n" + "─" * 78)
    logger.info(f" KEY FEATURES:")
    logger.info("  ✅ Technology-agnostic (works with ANY change management approach)")
    logger.info("  ✅ Standard/Normal/Emergency change types defined")
    logger.info("  ✅ Risk-based classification methodology")
    logger.info("  ✅ E-CAB procedures documented")
    logger.info("  ✅ Comprehensive evidence collection")
    logger.info("  ✅ Audit readiness assessment")
    logger.info("  ✅ Quarterly review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(f" NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Review Instructions & Legend sheet first")
    logger.info("  3. Populate Standard Changes Catalog with YOUR pre-approved changes")
    logger.info("  4. Document YOUR normal change criteria")
    logger.info("  5. Define YOUR emergency change triggers")
    logger.info("  6. Configure YOUR risk classification matrix")
    logger.info("  7. Document YOUR change calendar and blackout periods")
    logger.info("  8. Review Summary Dashboard for compliance status")
    logger.info("  9. Document evidence in Evidence Register")
    logger.info("  10. Obtain final approval via Approval Sign-Off")
    logger.info("\n PRO TIP:")
    logger.info("  If your emergency changes consistently stay below 5% of total changes")
    logger.info("  and your risk classifications accurately predict change outcomes,")
    logger.info("  you've moved beyond checkbox compliance to operational excellence.")
    logger.info("\n" + "=" * 78)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info("\n This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
