#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.32.5 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Consolidated Dashboard: Multi-Domain Assessment Consolidation & Executive Reporting

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dashboard requirements and reporting needs.

Key customization areas:
1. Dashboard layout and executive summary metrics
2. External workbook formula references (match your file structure)
3. Compliance threshold definitions (adapt to your risk tolerance)
4. Visualization preferences (charts, conditional formatting)
5. Integration with GRC platforms (optional)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates the consolidated compliance dashboard workbook that
aggregates data from all four change management assessment domains (A.8.32.1
through A.8.32.4) into a unified executive view.

**Purpose:**
Creates the master dashboard template with external workbook formula references
to automatically pull data from the four domain assessment workbooks, providing
executive visibility and audit-ready consolidated reporting.

**Relationship to consolidate_a832_dashboard.py:**
- generate_a832_5 (this script): Creates EMPTY dashboard template with formulas
- consolidate_a832_dashboard.py: POPULATES dashboard with assessment data values

Two methods for dashboard population:
1. Primary (recommended): External workbook formulas (auto-update when sources change)
2. Alternative: Data consolidation script (static snapshots)

**Dashboard Data Sources:**
1. Domain 1: Change Process Assessment (A.8.32.1)
   Source: ISMS_A_8_32_1_Change_Process_Assessment_YYYYMMDD.xlsx
   
2. Domain 2: Change Types & Categories (A.8.32.2)
   Source: ISMS_A_8_32_2_Change_Types_Categories_Assessment_YYYYMMDD.xlsx
   
3. Domain 3: Environment Separation (A.8.32.3)
   Source: ISMS_A_8_32_3_Environment_Separation_Assessment_YYYYMMDD.xlsx
   
4. Domain 4: Testing & Validation (A.8.32.4)
   Source: ISMS_A_8_32_4_Testing_Validation_Assessment_YYYYMMDD.xlsx

**Generated Dashboard Structure:**
1. Instructions & Legend - Dashboard usage guidance
2. Executive Dashboard - High-level metrics and KPIs (formulas from domains)
3. Gap Analysis - Consolidated gaps from all 4 domains
4. Evidence_Summary - Consolidated evidence registers
5. Risk_Matrix - Aggregated risk assessment
6. Compliance_Trend_Analysis - Historical compliance tracking
7. Audit_Findings - Consolidated audit findings
8. Remediation_Tracking - Gap remediation action tracking
9. Approval_Sign_Off - Executive approval workflow

**Key Features:**
- External workbook formula links (auto-update capability)
- Executive summary with aggregate compliance metrics
- Consolidated gap analysis from all domains
- Evidence summary with traceability
- Risk matrix aggregation
- Trend analysis for historical tracking
- Audit-ready consolidated reporting
- Multi-level approval workflow

**Integration:**
This dashboard consolidates data from A.8.32.1, A.8.32.2, A.8.32.3, and A.8.32.4
assessment workbooks using Excel external workbook references.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a832_5_compliance_dashboard.py

Output:
    File: ISMS_A_8_32_5_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Complete all four domain assessments (A.8.32.1 through A.8.32.4)
    2. Run normalize_assessment_files_a832.py to create normalized filenames
    3. Place dashboard and normalized assessments in same directory
    4. Open dashboard in Excel
    5. Click "Update Links" when prompted by Excel
    6. Review Executive Dashboard for overall compliance status
    7. Examine Gap Analysis for consolidated findings
    8. Use Remediation_Tracking for gap closure planning
    9. Obtain executive approvals via Approval_Sign_Off
    10. Archive for quarterly compliance reporting

Alternative Data Population:
    Instead of using external formulas, use consolidate_a832_dashboard.py to
    copy data values directly into dashboard (useful for static snapshots,
    archival, or when external links don't work).

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32
Dashboard Type:       Consolidated Multi-Domain Compliance Dashboard
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.5: Compliance Dashboard Implementation Guide
    - ISMS-IMP-A.8.32.1: Change Process Assessment (Input Domain 1)
    - ISMS-IMP-A.8.32.2: Change Types & Categories Assessment (Input Domain 2)
    - ISMS-IMP-A.8.32.3: Environment Separation Assessment (Input Domain 3)
    - ISMS-IMP-A.8.32.4: Testing & Validation Assessment (Input Domain 4)

Related Scripts:
    - generate_a832_1_change_process.py (generates input)
    - generate_a832_2_change_types.py (generates input)
    - generate_a832_3_environment_separation.py (generates input)
    - generate_a832_4_testing_validation.py (generates input)
    - consolidate_a832_dashboard.py (alternative population method)
    - normalize_assessment_files_a832.py (file preparation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements dashboard template with external workbook references
    - Supports consolidated reporting from all 4 assessment domains
    - Generates audit-ready compliance dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**External Workbook Links:**
Dashboard uses Excel external workbook references like:
='[ISMS-IMP-A.8.32.1.xlsx]Summary_Dashboard'!$B$15

These links require:
- Dashboard and source files in same directory
- Normalized filenames (run normalize_assessment_files_a832.py)
- "Update Links" when opening dashboard

If external links don't work (Excel security, SharePoint issues), use
consolidate_a832_dashboard.py to copy data values instead.

**Dashboard Maintenance:**
- Quarterly: Re-run assessments, update dashboard links
- Semi-annually: Review trend analysis and improvement trajectory
- Annually: Complete reassessment and executive presentation
- Ad-hoc: When significant changes to change management processes

**Audit Considerations:**
This dashboard is the primary artifact for ISO 27001:2022 Control A.8.32
compliance audits. Auditors expect:
- Consolidated view of all 4 domains
- Traceability to source assessments
- Evidence of executive oversight
- Gap remediation tracking

**Data Protection:**
Dashboard contains executive summary of organizational change management
maturity including compliance gaps and remediation priorities. Handle with
appropriate confidentiality.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.32.5"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.8.32"
CONTROL_NAME = "Change Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
CLOCK = '\U0001F552'  # 🕒 Clock
WRENCH = '\U0001F527' # 🔧 Wrench
ROCKET = '\U0001F680' # 🚀 Rocket
GEAR = '\u2699'       # ⚙️  Gear
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.32.5 specification (8 sheets)
    sheets = [
        "Executive Dashboard",
        "Gap Analysis",
        "Risk Register",
        "Remediation Roadmap",
        "KPIs Metrics",
        "Evidence Register",
        "Audit Readiness",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "critical_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_yellow": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: EXECUTIVE DASHBOARD
# ============================================================================

def create_executive_dashboard(ws, styles):
    """Create Executive Dashboard with external workbook links and KPIs."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "ISMS-IMP-A.8.32.5  Change Management Compliance Summary Dashboard\n"
        "ISO/IEC 27001:2022 - Control A.8.32: Change Management - Executive Overview"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # ---------- DOCUMENT INFORMATION ----------
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.32.5"),
        ("Report Type", "Compliance Summary Dashboard"),
        ("Related Policy", "ISMS-POL-A.8.32"),
        ("Version", "1.0"),
        ("Report Date", ""),
        ("Reporting Period", ""),
        ("Prepared By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
        ("Last Updated", "=TODAY()"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        elif "=" in str(value):
            ws[f"B{row}"].font = Font(color="0000FF")
        row += 1

    # ---------- OVERALL COMPLIANCE STATUS ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "OVERALL CHANGE MANAGEMENT COMPLIANCE STATUS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    # Compliance Scorecard
    row += 1
    scorecard_headers = ["Metric", "Score", "Target", "Status"]
    for col_idx, header in enumerate(scorecard_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    scorecard_start = row + 1
    scorecard_metrics = [
        ("Overall Compliance Rate", "=AVERAGE(C25:C28)", "95%", ""),
        ("Critical Gaps", "=COUNTIFS(Gap Analysis!I8:I207,\"Critical\",Gap Analysis!R8:R207,\"<>Closed\")", "0", ""),
        ("High-Risk Items", "=COUNTIFS(Risk Register!N20:N119,\">=15\",Risk Register!Q20:Q119,\"<>Closed\")", "5", ""),
        ("Remediation Progress", "=COUNTIF(Remediation Roadmap!H37:H236,\"Completed\")/COUNTA(Remediation Roadmap!H37:H236)*100&\"%\"", "80%", ""),
    ]

    row += 1
    for metric, formula, target, status in scorecard_metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        ws.cell(row=row, column=2, value=formula).font = Font(bold=True, color="0000FF")
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    # Add dropdown for status column
    dv_status = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(f"D{scorecard_start}:D{row-1}")

    # ---------- DOMAIN COMPLIANCE SCORES ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "DOMAIN COMPLIANCE SCORES (from Source Workbooks)"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    domain_headers = ["#", "Assessment Domain", "Compliance Score", "Status", "Critical Gaps", "Evidence Count", "Last Updated", "Assessor"]
    for col_idx, header in enumerate(domain_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Domain rows with EXTERNAL WORKBOOK LINKS
    # This is where the magic happens - pulling data from normalized source files!
    domain_data = [
        (
            "1",
            "Change Process",
            "='[ISMS-IMP-A.8.32.1.xlsx]Gap Analysis'!B5",  # Link to score cell
            "='[ISMS-IMP-A.8.32.1.xlsx]Gap Analysis'!C5",  # Link to status
            "=COUNTIF('[ISMS-IMP-A.8.32.1.xlsx]Gap Analysis'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.32.1.xlsx]Evidence Register'!A10:A209)",
            "='[ISMS-IMP-A.8.32.1.xlsx]Instructions_Legend'!B5",  # Assessment date
            "='[ISMS-IMP-A.8.32.1.xlsx]Instructions_Legend'!B6",  # Completed by
        ),
        (
            "2",
            "Change Types & Categories",
            "='[ISMS-IMP-A.8.32.2.xlsx]Gap_Identification'!B5",
            "='[ISMS-IMP-A.8.32.2.xlsx]Gap_Identification'!C5",
            "=COUNTIF('[ISMS-IMP-A.8.32.2.xlsx]Gap_Identification'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.32.2.xlsx]Evidence Register'!A10:A209)",
            "='[ISMS-IMP-A.8.32.2.xlsx]Instructions_Legend'!B5",
            "='[ISMS-IMP-A.8.32.2.xlsx]Instructions_Legend'!B6",
        ),
        (
            "3",
            "Environment Separation",
            "='[ISMS-IMP-A.8.32.3.xlsx]Gap Analysis'!B5",
            "='[ISMS-IMP-A.8.32.3.xlsx]Gap Analysis'!C5",
            "=COUNTIF('[ISMS-IMP-A.8.32.3.xlsx]Gap Analysis'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.32.3.xlsx]Evidence Register'!A10:A209)",
            "='[ISMS-IMP-A.8.32.3.xlsx]Instructions_Legend'!B5",
            "='[ISMS-IMP-A.8.32.3.xlsx]Instructions_Legend'!B6",
        ),
        (
            "4",
            "Testing & Validation",
            "='[ISMS-IMP-A.8.32.4.xlsx]Gap Analysis'!B5",
            "='[ISMS-IMP-A.8.32.4.xlsx]Gap Analysis'!C5",
            "=COUNTIF('[ISMS-IMP-A.8.32.4.xlsx]Gap Analysis'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.32.4.xlsx]Evidence Register'!A10:A209)",
            "='[ISMS-IMP-A.8.32.4.xlsx]Instructions_Legend'!B5",
            "='[ISMS-IMP-A.8.32.4.xlsx]Instructions_Legend'!B6",
        ),
    ]

    row += 1
    for domain in domain_data:
        for col_idx, value in enumerate(domain, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx == 3:  # Compliance score
                cell.number_format = "0.0%"
                cell.fill = styles["formula_cell"]["fill"]
            elif col_idx in [4, 5, 6, 7, 8]:  # Formula columns
                cell.fill = styles["formula_cell"]["fill"]
            if "=" in str(value):
                cell.font = Font(color="0000FF", italic=True, size=9)
        row += 1

    # ---------- KEY PERFORMANCE INDICATORS ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY PERFORMANCE INDICATORS (Consolidated)"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    kpi_headers = ["Category", "Metric", "Current Value", "Target", "Status", "Trend", "Notes"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    kpi_metrics = [
        ("Infrastructure", "Filtering Solutions Deployed", "=COUNTA('[ISMS-IMP-A.8.32.1.xlsx]Solution_Details_Template'!B8:B27)", "1", "", "", "From A.8.32.1"),
        ("Infrastructure", "Solutions Fully Implemented", "=COUNTIF('[ISMS-IMP-A.8.32.1.xlsx]Solution_Details_Template'!I8:I27,\"Deployed\")/COUNTA('[ISMS-IMP-A.8.32.1.xlsx]Solution_Details_Template'!B8:B27)", "100%", "", "", ""),
        ("Coverage", "Network Segments Covered", "=COUNTIF('[ISMS-IMP-A.8.32.2.xlsx]Coverage_Matrix'!D10:D109,\" Covered\")/COUNTA('[ISMS-IMP-A.8.32.2.xlsx]Network_Segment_Inventory'!A10:A59)", "100%", "", "", "From A.8.32.2"),
        ("Coverage", "Critical Bypass Risks", "=COUNTIF('[ISMS-IMP-A.8.32.2.xlsx]Gap_Identification'!F10:F109,\"Critical\")", "0", "", "", ""),
        ("Policy", "Threat Categories Blocked", "=COUNTIF('[ISMS-IMP-A.8.32.3.xlsx]Threat_Protection'!E10:E59,\"Blocked\")/COUNTA('[ISMS-IMP-A.8.32.3.xlsx]Threat_Protection'!A10:A59)", "95%", "", "", "From A.8.32.3"),
        ("Policy", "Active Exceptions", "=COUNTIF('[ISMS-IMP-A.8.32.3.xlsx]Policy_Exceptions'!G10:G109,\"Active\")", "<10", "", "", ""),
        ("Monitoring", "Alert Rules Configured", "=COUNTA('[ISMS-IMP-A.8.32.4.xlsx]Alert_Configuration'!A10:A109)", "20", "", "", "From A.8.32.4"),
        ("Monitoring", "SLA Compliance Rate", "=COUNTIF('[ISMS-IMP-A.8.32.4.xlsx]Incident_Response'!K10:K109,\"Met\")/COUNTA('[ISMS-IMP-A.8.32.4.xlsx]Incident_Response'!K10:K109)", "95%", "", "", ""),
        ("Response", "Open Critical Incidents", "=COUNTIFS('[ISMS-IMP-A.8.32.4.xlsx]Incident_Response'!D10:D109,\"Critical\",'[ISMS-IMP-A.8.32.4.xlsx]Incident_Response'!I10:I109,\"Open\")", "0", "", "", ""),
        ("Response", "False Positive Rate", "=COUNTIF('[ISMS-IMP-A.8.32.4.xlsx]False_Positive_Management'!H10:H109,\"Confirmed_FP\")/COUNTA('[ISMS-IMP-A.8.32.4.xlsx]Blocked_Events_Analysis'!A10:A109)", "<1%", "", "", ""),
    ]

    row += 1
    kpi_start_row = row
    for category, metric, formula, target, status, trend, notes in kpi_metrics:
        ws.cell(row=row, column=1, value=category).border = styles["border"]
        ws.cell(row=row, column=2, value=metric).border = styles["border"]
        ws.cell(row=row, column=3, value=formula).border = styles["border"]
        ws.cell(row=row, column=3).font = Font(color="0000FF", italic=True, size=9)
        ws.cell(row=row, column=3).fill = styles["formula_cell"]["fill"]
        ws.cell(row=row, column=4, value=target).border = styles["border"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).border = styles["border"]
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6).border = styles["border"]
        ws.cell(row=row, column=7, value=notes).border = styles["border"]
        ws.cell(row=row, column=7).font = Font(italic=True, size=9, color="666666")
        row += 1

    # Add dropdowns
    dv_status_kpi = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=False)
    ws.add_data_validation(dv_status_kpi)
    dv_status_kpi.add(f"E{kpi_start_row}:E{row-1}")

    dv_trend = DataValidation(type="list", formula1='"Improving,Stable,Degrading"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add(f"F{kpi_start_row}:F{row-1}")

    # ---------- FOOTER NOTES ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = " DASHBOARD NOTES"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    notes = [
        " This dashboard consolidates data from 4 assessment workbooks via external links",
        " Blue italic formulas pull live data from normalized source files",
        " Click 'Update Links' when opening to refresh all data",
        " Yellow cells are for manual status/trend assessment",
        " Place this dashboard in same folder as ISMS-IMP-A.8.32.[1-4].xlsx files",
        " Review quarterly or after significant changes to filtering infrastructure",
    ]

    row += 1
    for note in notes:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = note
        ws[f"A{row}"].font = Font(size=9, italic=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = " \"Evidence > Theater\" - This is Systems Engineering, not Cargo Cult Compliance"
    ws[f"A{row}"].font = Font(bold=True, italic=True, size=10, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A4"

# SECTION 3: GAP ANALYSIS
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create consolidated Gap Analysis sheet from all 4 assessments."""
    
    ws.merge_cells("A1:S1")
    ws["A1"] = "CONSOLIDATED GAP ANALYSIS - A.8.32 CHANGE MANAGEMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:S2")
    ws["A2"] = "Aggregated gaps from all 4 assessment workbooks"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = [
        "Gap ID", "Source Assessment", "Gap Category", "Gap Description", 
        "Current State", "Target State", "Impact", "Likelihood", "Risk Score",
        "Priority", "Remediation Action", "Owner", "Target Date", 
        "Status", "Progress %", "Cost Estimate", "Dependencies", 
        "Verification Method", "Status Notes"
    ]
    
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Instructions
    row = 5
    ws.merge_cells(f"A{row}:S{row}")
    ws[f"A{row}"] = (
        "INSTRUCTIONS: Consolidate gaps from all 4 source workbooks. "
        "Each gap should reference its source assessment. "
        "Use this sheet for executive-level gap tracking across all change management domains."
    )
    ws[f"A{row}"].font = Font(italic=True, size=9, color="666666")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30

    # Example gap entries (50 rows for consolidation)
    row = 6
    example_gaps = [
        ("A.8.32.1-GAP-001", "Change Process", "Technology", "", "", "", "High", "Medium", "=G7*H7", "High", "", "", "", "", "", "", "", "", ""),
        ("A.8.32.2-GAP-001", "Change Types & Categories", "Coverage", "", "", "", "Critical", "High", "=G8*H8", "Critical", "", "", "", "", "", "", "", "", ""),
        ("A.8.32.3-GAP-001", "Environment Separation", "Policy", "", "", "", "Medium", "Low", "=G9*H9", "Medium", "", "", "", "", "", "", "", "", ""),
        ("A.8.32.4-GAP-001", "Testing & Validation", "Monitoring", "", "", "", "High", "Medium", "=G10*H10", "High", "", "", "", "", "", "", "", "", ""),
    ]

    for gap_data in example_gaps:
        for col_idx, value in enumerate(gap_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            if value == "":
                cell.fill = styles["input_cell"]["fill"]
            elif "=" in str(value):
                cell.font = Font(color="0000FF")
        row += 1

    # Add empty rows for additional gaps (total 200 rows)
    for _ in range(46):
        for col_idx in range(1, 20):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
        row += 1

    # Add dropdowns
    dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add(f"J7:J{row-1}")

    dv_status = DataValidation(type="list", formula1='"Open,In_Progress,Resolved,Accepted,Deferred,Closed"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"N7:N{row-1}")

    # Column widths
    widths = [15, 25, 20, 40, 30, 30, 10, 12, 12, 12, 35, 20, 15, 15, 12, 15, 25, 25, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 4: RISK REGISTER
# ============================================================================

def create_risk_register(ws, styles):
    """Create consolidated Risk Register."""
    
    ws.merge_cells("A1:Q1")
    ws["A1"] = "RISK REGISTER - A.8.32 CHANGE MANAGEMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Risk matrix reference
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "RISK SCORING MATRIX"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    row += 1
    ws[f"A{row}"] = "Impact / Likelihood"
    ws[f"B{row}"] = "Low (1)"
    ws[f"C{row}"] = "Medium (2)"
    ws[f"D{row}"] = "High (3)"
    ws[f"E{row}"] = "Critical (4)"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True, size=9)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col}{row}"].border = styles["border"]

    risk_matrix = [
        ("Critical (4)", "4", "8", "12", "16"),
        ("High (3)", "3", "6", "9", "12"),
        ("Medium (2)", "2", "4", "6", "8"),
        ("Low (1)", "1", "2", "3", "4"),
    ]

    row += 1
    for risk_row in risk_matrix:
        ws[f"A{row}"] = risk_row[0]
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].border = styles["border"]
        for col_idx, val in enumerate(risk_row[1:], start=2):
            cell = ws.cell(row=row, column=col_idx, value=int(val))
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="center")
            # Color code by score
            score = int(val)
            if score >= 12:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif score >= 6:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

    # Risk register table
    row += 2
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "IDENTIFIED RISKS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    headers = [
        "Risk ID", "Source", "Risk Category", "Risk Description", 
        "Threat Source", "Vulnerability", "Impact", "Likelihood", 
        "Inherent Risk", "Controls", "Control Effectiveness", 
        "Residual Impact", "Residual Likelihood", "Residual Risk", 
        "Treatment", "Owner", "Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Add 100 risk entry rows
    row += 1
    for i in range(100):
        for col_idx in range(1, 18):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
            
            # Add formulas for risk calculations
            if col_idx == 9:  # Inherent Risk
                cell.value = f"=G{row}*H{row}"
                cell.font = Font(color="0000FF")
            elif col_idx == 14:  # Residual Risk
                cell.value = f"=L{row}*M{row}"
                cell.font = Font(color="0000FF")
        row += 1

    # Add dropdowns
    # Impact ranges split for proper validation
    dv_impact = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    ws.add_data_validation(dv_impact)
    dv_impact.add(f"G{row-100}:G{row-1}")
    dv_impact.add(f"L{row-100}:L{row-1}")

    # Likelihood ranges split for proper validation
    dv_likelihood = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    ws.add_data_validation(dv_likelihood)
    dv_likelihood.add(f"H{row-100}:H{row-1}")
    dv_likelihood.add(f"M{row-100}:M{row-1}")

    dv_treatment = DataValidation(type="list", formula1='"Accept,Mitigate,Transfer,Avoid"', allow_blank=True)
    ws.add_data_validation(dv_treatment)
    dv_treatment.add(f"O{row-100}:O{row-1}")

    dv_status = DataValidation(type="list", formula1='"Open,Monitoring,Resolved,Closed"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"Q{row-100}:Q{row-1}")

    # Column widths
    widths = [15, 20, 20, 40, 25, 25, 10, 12, 12, 30, 15, 10, 12, 12, 15, 20, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A20"


# ============================================================================
# SECTION 5: REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap(ws, styles):
    """Create Remediation Roadmap with timeline tracking."""
    
    ws.merge_cells("A1:P1")
    ws["A1"] = "REMEDIATION ROADMAP - A.8.32 CHANGE MANAGEMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:P2")
    ws["A2"] = "Action plan for closing identified gaps and reducing risks"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Roadmap summary
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "REMEDIATION SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    summary_metrics = [
        ("Total Action Items", "=COUNTA(A37:A236)"),
        ("Completed", "=COUNTIF(H37:H236,\"Completed\")"),
        ("In Progress", "=COUNTIF(H37:H236,\"In Progress\")"),
        ("Not Started", "=COUNTIF(H37:H236,\"Not Started\")"),
        ("Overdue", "=COUNTIFS(H37:H236,\"<>Completed\",E37:E236,\"<\"&TODAY())"),
        ("Overall Progress", "=COUNTIF(H37:H236,\"Completed\")/COUNTA(A37:A236)"),
    ]

    row += 1
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(color="0000FF", bold=True)
        ws[f"B{row}"].border = styles["border"]
        if "Progress" in label:
            ws[f"B{row}"].number_format = "0.0%"
        row += 1

    # Timeline phases
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "IMPLEMENTATION PHASES"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    phases = [
        ("Phase 1: Critical (0-30 days)", "Immediate action required", "Critical gaps, security risks"),
        ("Phase 2: High (30-90 days)", "Quick wins, high impact", "Infrastructure, coverage gaps"),
        ("Phase 3: Medium (90-180 days)", "Process improvements", "Policy, monitoring enhancements"),
        ("Phase 4: Low (180-365 days)", "Optimization", "Fine-tuning, documentation"),
    ]

    row += 1
    for phase, description, focus in phases:
        ws[f"A{row}"] = phase
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = description
        ws[f"B{row}"].border = styles["border"]
        ws.merge_cells(f"C{row}:F{row}")
        ws[f"C{row}"] = focus
        ws[f"C{row}"].border = styles["border"]
        ws[f"C{row}"].font = Font(italic=True, size=9)
        row += 1

    # Action item table
    row += 2
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "ACTION ITEMS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    headers = [
        "Action ID", "Phase", "Related Gap/Risk", "Action Description",
        "Target Date", "Assigned Owner", "Department", "Status",
        "Progress %", "Actual Completion", "Budget", "Dependencies",
        "Success Criteria", "Verification", "Notes", "Last Updated"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Add 200 action item rows
    row += 1
    for i in range(200):
        for col_idx in range(1, 17):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
        row += 1

    # Add dropdowns
    dv_phase = DataValidation(type="list", formula1='"Phase 1,Phase 2,Phase 3,Phase 4"', allow_blank=True)
    ws.add_data_validation(dv_phase)
    dv_phase.add(f"B37:B236")

    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked,Cancelled"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"H37:H236")

    # Column widths
    widths = [15, 15, 20, 40, 15, 20, 15, 15, 12, 15, 15, 25, 30, 20, 30, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A37"

# SECTION 6: KPIs & METRICS
# ============================================================================

def create_kpis_metrics(ws, styles):
    """Create KPIs & Metrics tracking sheet."""
    
    ws.merge_cells("A1:M1")
    ws["A1"] = "KPIs & METRICS - A.8.32 CHANGE MANAGEMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # KPI Categories
    categories = [
        {
            "name": "INFRASTRUCTURE METRICS",
            "kpis": [
                ("Total Filtering Solutions", "Count", "", "1", ""),
                ("Deployment Coverage", "%", "", "100%", ""),
                ("Solution Availability", "%", "", "99%", ""),
                ("Performance - Latency", "ms", "", "100ms", ""),
                ("Licensing Compliance", "%", "", "100%", ""),
            ]
        },
        {
            "name": "COVERAGE METRICS",
            "kpis": [
                ("Network Segments Covered", "%", "", "100%", ""),
                ("Users Protected", "Count", "", "All", ""),
                ("Devices Protected", "Count", "", "All", ""),
                ("Bypass Vulnerabilities", "Count", "", "0", ""),
                ("Coverage Verification Rate", "%", "", "100%", ""),
            ]
        },
        {
            "name": "POLICY METRICS",
            "kpis": [
                ("Threat Categories Blocked", "%", "", "95%", ""),
                ("Content Categories Configured", "Count", "", "30", ""),
                ("Active Policy Exceptions", "Count", "", "10", ""),
                ("Exception Review Compliance", "%", "", "100%", ""),
                ("Policy Update Frequency", "Days", "", "90", ""),
            ]
        },
        {
            "name": "MONITORING METRICS",
            "kpis": [
                ("Log Collection Rate", "%", "", "100%", ""),
                ("Alert Rules Active", "Count", "", "20", ""),
                ("Alert Response Time", "Minutes", "", "30", ""),
                ("Dashboard Availability", "%", "", "99%", ""),
                ("SIEM Integration Health", "%", "", "100%", ""),
            ]
        },
        {
            "name": "INCIDENT METRICS",
            "kpis": [
                ("Open Critical Incidents", "Count", "", "0", ""),
                ("Mean Time to Detect (MTTD)", "Minutes", "", "15", ""),
                ("Mean Time to Respond (MTTR)", "Hours", "", "4", ""),
                ("SLA Compliance Rate", "%", "", "95%", ""),
                ("False Positive Rate", "%", "", "1%", ""),
            ]
        },
        {
            "name": "COMPLIANCE METRICS",
            "kpis": [
                ("Overall Compliance Score", "%", "", "95%", ""),
                ("Critical Gaps Open", "Count", "", "0", ""),
                ("Evidence Collection Rate", "%", "", "100%", ""),
                ("Audit Readiness", "Status", "", "Ready", ""),
                ("Policy Review Currency", "Days", "", "90", ""),
            ]
        },
    ]

    row = 3
    for category in categories:
        # Category header
        ws.merge_cells(f"A{row}:M{row}")
        ws[f"A{row}"] = category["name"]
        ws[f"A{row}"].font = styles["section_header"]["font"]
        ws[f"A{row}"].fill = styles["section_header"]["fill"]
        ws[f"A{row}"].alignment = styles["section_header"]["alignment"]
        
        row += 1
        # Column headers
        headers = ["KPI Name", "Unit", "Current Value", "Target", "Status", "Trend", 
                   "Last Month", "YTD Average", "Data Source", "Update Frequency", 
                   "Owner", "Collection Method", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = styles["column_header"]["font"]
            cell.fill = styles["column_header"]["fill"]
            cell.alignment = styles["column_header"]["alignment"]
            cell.border = styles["border"]
        
        row += 1
        # KPI rows
        for kpi_name, unit, current, target, status in category["kpis"]:
            ws[f"A{row}"] = kpi_name
            ws[f"A{row}"].border = styles["border"]
            ws[f"B{row}"] = unit
            ws[f"B{row}"].border = styles["border"]
            ws[f"C{row}"].fill = styles["input_cell"]["fill"]
            ws[f"C{row}"].border = styles["border"]
            ws[f"D{row}"] = target
            ws[f"D{row}"].border = styles["border"]
            for col in ["E", "F", "G", "H", "I", "J", "K", "L", "M"]:
                ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
                ws[f"{col}{row}"].border = styles["border"]
            row += 1
        
        row += 1  # Blank row between categories

    # Add dropdowns
    dv_status = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("E4:E100")

    dv_trend = DataValidation(type="list", formula1='"Improving,Stable,Degrading,New"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add("F4:F100")

    dv_frequency = DataValidation(type="list", formula1='"Real-Time,Hourly,Daily,Weekly,Monthly"', allow_blank=True)
    ws.add_data_validation(dv_frequency)
    dv_frequency.add("J4:J100")

    # Column widths
    widths = [30, 10, 15, 15, 12, 12, 15, 15, 25, 18, 20, 20, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create standard Evidence Register (8 columns, 100 rows)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header ---
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # --- Subtitle ---
    ws.merge_cells("A2:H2")
    ws["A2"] = "Centralised evidence tracking for audit and compliance verification"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Column headers (row 4) ---
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # --- 100 data rows (5-104) ---
    for i in range(1, 101):
        row = i + 4
        # Col A — Evidence ID (gray font, no input fill)
        cell_a = ws.cell(row=row, column=1, value=f"EV-{i:03d}")
        cell_a.font = Font(name="Calibri", color="808080")
        cell_a.alignment = Alignment(horizontal="center", vertical="center")
        # Cols B-H — yellow input
        for col_idx in range(2, 9):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # --- Dropdowns ---
    dv_type = DataValidation(
        type="list",
        formula1='"Screenshot,Configuration Export,Log Sample,Report,Procedure,Email,Audit Report,Test Results,Policy Document,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C104")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending,Not Verified,Expired"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H104")

    # --- Column widths ---
    for col_letter, w in zip("ABCDEFGH", [15, 25, 22, 40, 45, 16, 20, 22]):
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 9: AUDIT READINESS SHEET
# ============================================================================

def create_audit_readiness(ws, styles):
    """Create Audit Readiness sheet for audit preparation assessment."""

    ws.merge_cells("A1:G1")
    ws["A1"] = "AUDIT READINESS ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Assessment of readiness for internal and external audits"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["Audit Area", "Evidence Available", "Gaps Identified", "Remediation Status", "Target Date", "Owner", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
    row += 1

    areas = [
        "Change Process Documentation",
        "Change Type Classification",
        "Environment Separation",
        "Testing & Validation",
        "Evidence Collection",
        "Risk Management",
        "Compliance Metrics",
    ]

    for area in areas:
        ws.cell(row=row, column=1, value=area)
        for col_idx in range(2, 8):
            cell = ws.cell(row=row, column=col_idx, value="")
            cell.fill = styles["input_cell"]["fill"]
        row += 1

    widths = [30, 20, 25, 20, 15, 18, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 9.5: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff_sheet(ws, styles):
    """Create standard Approval Sign-Off sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # --- ASSESSMENT SUMMARY ---
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    summary_fields = [
        ("Document ID:", DOCUMENT_ID, False),
        ("Assessment:", WORKBOOK_NAME, False),
        ("Control Reference:", CONTROL_REF, False),
        ("Assessment Status:", "", True),
        ("Overall Rating:", "", True),
    ]
    row = 4
    for label, value, is_input in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if is_input:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Assessment Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_status)
    dv_status.add("B7")

    # --- Helper for approver sections ---
    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="center", vertical="center")
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{r}:E{r}")
            ws[f"B{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{r}"].border = border
            r += 1
        return r

    # --- COMPLETED BY ---
    row += 1
    row = _approver_section(row, "COMPLETED BY", "4472C4")

    # --- REVIEWED BY ---
    row += 1
    row = _approver_section(row, "REVIEWED BY", "4472C4")

    # --- APPROVED BY — CISO ---
    row += 1
    row = _approver_section(row, "APPROVED BY \u2014 CISO", "003366")

    # --- FINAL DECISION ---
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(f"B{row}")

    # --- NEXT REVIEW DETAILS ---
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Next Review Date:", "Review Frequency:", "Scheduled Reviewer:"]:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        row += 1

    # --- Column widths ---
    for col_letter, w in zip("ABCDE", [32, 25, 20, 20, 20]):
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 10: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.32.5 - Compliance Summary Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.32: Change Management")
    logger.info("=" * 80)
    logger.info("\n Systems Engineering Approach - Consolidated Oversight Dashboard")
    logger.info("   This dashboard pulls live data from 4 assessment workbooks\n")

    wb = create_workbook()
    styles = setup_styles()

    logger.info("\n[1/8] Creating Executive Dashboard sheet...")
    create_executive_dashboard(wb["Executive Dashboard"], styles)

    logger.info("[2/8] Creating Gap Analysis sheet...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[3/8] Creating Risk Register sheet...")
    create_risk_register(wb["Risk Register"], styles)

    logger.info("[4/8] Creating Remediation Roadmap sheet...")
    create_remediation_roadmap(wb["Remediation Roadmap"], styles)

    logger.info("[5/8] Creating KPIs & Metrics sheet...")
    create_kpis_metrics(wb["KPIs Metrics"], styles)

    logger.info("[6/8] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[7/8] Creating Audit Readiness sheet...")
    create_audit_readiness(wb["Audit Readiness"], styles)

    logger.info("[8/8] Creating Approval Sign-Off sheet...")
    create_approval_signoff_sheet(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.8.32.5_Compliance_Summary_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info(f"\n SUCCESS: {filename}")
    logger.info("\n" + "=" * 80)
    logger.info("COMPLIANCE SUMMARY DASHBOARD - COMPLETE")
    logger.info("=" * 80)
    logger.info("\nWorkbook Structure (8 Sheets):")
    logger.info("  1. Executive Dashboard - Overall compliance with external links")
    logger.info("  2. Gap Analysis - Consolidated gaps from all 4 assessments")
    logger.info("  3. Risk Register - Risk entries with inherent/residual scoring")
    logger.info("  4. Remediation Roadmap - Remediation items with timeline tracking")
    logger.info("  5. KPIs Metrics - KPIs across 6 categories")
    logger.info("  6. Evidence Register - 100 evidence entries (standard format)")
    logger.info("  7. Audit Readiness - Audit preparation assessment")
    logger.info("  8. Approval Sign-Off - Standard 3-section approval workflow")
    logger.info("=" * 80 + "\n")


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT - COMPLETE COMPLIANCE SUMMARY DASHBOARD GENERATOR
# ============================================================================

# =============================================================================
# QA_VERIFIED: 2026-02-10
# QA_STATUS: PASSED - STANDARDISATION COMPLETE
# QA_TOOL: Claude Code Standardisation
# CHANGES: tab names (underscores→spaces), Evidence Register (standard 8-col/100-row),
#   Approval Sign-Off (standard 3-section), dead code removed (action_items/audit_log/
#   legacy approval_signoff), Related Policy fix, logger output updated
# =============================================================================
