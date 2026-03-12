#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.32.1 - Change Process Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Assessment Domain 1 of 4: Change Process Workflow & Management Procedures

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific change management processes and assessment
requirements.

Key customization areas:
1. Sheet names and structures (match your control's assessment needs)
2. Data validation rules (specific to your change management requirements)
3. Compliance scoring logic (adapt to your risk profile)
4. Output formatting (match your assessment workbook structure)
5. Change process stages (customize to your workflow)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
change management processes and workflows against ISO 27001:2022 Control A.8.32
requirements, supporting evidence-based validation of change management
procedures.

**Purpose:**
Enables systematic assessment of change management processes including workflow
stages, approval authorities, communication procedures, documentation requirements,
and tool capabilities against ISO 27001:2022 Control A.8.32 requirements.

**Assessment Scope:**
- Change process workflow definition and stages
- Approval authority matrix and decision rights
- Stakeholder communication procedures and timing
- Documentation requirements and retention
- Change management tool capabilities and integration
- Process compliance verification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure (11 sheets per IMP specification):**
1. Instructions & Legend - Assessment guidance and standards
2. Change Process Workflow - Process stages and procedures
3. Approval Authority Matrix - Authority levels and thresholds
4. CAB Operations - Change Advisory Board composition and operation
5. Communication - Stakeholder notification methods
6. Documentation Records - Record-keeping standards
7. Tool Capabilities - Tool inventory and capabilities
8. Metrics KPIs - Tracked metrics and reporting
9. Evidence Register - Audit evidence tracking
10. Summary Dashboard - Compliance metrics and analytics
11. Approval Sign-Off - Stakeholder approval workflow

**Key Features:**
- Technology-agnostic assessment (works with any change management tool)
- Data validation with dropdown lists for consistency
- Conditional formatting for visual status indicators
- Automated compliance calculations
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
consolidates data from all four change management assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a832_1_change_process.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a832_1_change_process.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a832_1_change_process.py --date 20260127

Output:
    File: ISMS_A_8_32_1_Change_Process_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review Instructions & Legend sheet for assessment guidance
    2. Document your organisation's change process workflow stages
    3. Complete approval authority matrix with your roles and thresholds
    4. Define your stakeholder communication procedures
    5. Document your record-keeping and retention requirements
    6. Inventory your change management tools and capabilities
    7. Review Summary Dashboard for compliance metrics
    8. Collect and link audit evidence in Evidence Register
    9. Obtain stakeholder approvals via Approval Sign-Off

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32
Assessment Domain:    1 of 4 (Change Process Workflow & Management)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.1: Change Process Assessment Implementation Guide
    - ISMS-IMP-A.8.32.2: Change Types & Categories Assessment (Domain 2)
    - ISMS-IMP-A.8.32.3: Environment Separation Assessment (Domain 3)
    - ISMS-IMP-A.8.32.4: Testing & Validation Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.32.1 specification
    - Supports comprehensive change process evaluation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Technology-Agnostic Design:**
This assessment framework is intentionally tool-independent. Whether your
organisation uses ServiceNow, Jira, BMC Remedy, custom tools, spreadsheets,
or even paper-based processes, this assessment evaluates your PROCESS and
COMPLIANCE, not your tool brands.

**Process vs. Tool Assessment:**
The assessment focuses on:
- Is your change process defined and documented?
- Are approval authorities clear and appropriate?
- Are stakeholders notified appropriately?
- Is documentation complete and retained?
- Does your tool support (not dictate) your process?

NOT:
- Which vendor's tool you use
- How much you paid for your tool
- Whether you have the "latest and greatest" platform

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of your change process, not your tool choice.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Change process workflows and procedures
- Approval authorities and escalation paths
- Tool capabilities and limitations
- Identified compliance gaps

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for process changes and tool updates
- Semi-annually: Review approval authorities
- Annually: Complete reassessment of all domains
- Ad-hoc: When significant process or organisational changes occur

**Quality Assurance:**
Have change management practitioners and security personnel validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Customize assessment criteria to include regulatory-specific requirements:
- Financial services: SOX change management controls
- Healthcare: HIPAA security rule change management
- Government: Jurisdiction-specific change control requirements
- PCI DSS v4.0.1: Change management for cardholder data environments

Adapt dropdown values and assessment questions to your regulatory context.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from pathlib import Path
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.32.1"
WORKBOOK_NAME = "Change Process Assessment"
CONTROL_ID = "A.8.32"
CONTROL_NAME = "Change Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠  Warning sign
GEAR = '\u2699'       # ⚙  Gear
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP-A.8.32.1 spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches ISMS-IMP-A.8.32.1 specification (11 sheets)
    sheets = [
        "Instructions & Legend",
        "Change Process Workflow",
        "Approval Authority Matrix",
        "CAB Operations",
        "Communication",
        "Documentation Records",
        "Tool Capabilities",
        "Metrics KPIs",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_implemented": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notimplemented": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        },
        "compliance_full": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "compliance_substantial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "compliance_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "compliance_non": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    
    CUSTOMIZE: Adapt dropdown values for change management context.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Implemented,⚠ Partial,❌ Not Implemented, Planned,N/A"',
            allow_blank=False
        ),
        'automation_level': DataValidation(
            type="list",
            formula1='"Automated,Semi-Automated,Manual"',
            allow_blank=False
        ),
        'change_type': DataValidation(
            type="list",
            formula1='"Standard,Normal,Emergency"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Low,Medium,High,Critical"',
            allow_blank=False
        ),
        'impact_level': DataValidation(
            type="list",
            formula1='"Low,Medium,High,Critical"',
            allow_blank=False
        ),
        'cab_required': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'approval_method': DataValidation(
            type="list",
            formula1='"Virtual Meeting,Phone,Email Chain,Instant Messaging,Other"',
            allow_blank=False
        ),
        'notification_trigger': DataValidation(
            type="list",
            formula1='"Submission,Approval,24h Before,At Implementation,Completion,All"',
            allow_blank=False
        ),
        'communication_method': DataValidation(
            type="list",
            formula1='"Email,Portal,IM,SMS,All"',
            allow_blank=False
        ),
        'audience_reach': DataValidation(
            type="list",
            formula1='"All Users,IT Staff,Management,Selective"',
            allow_blank=False
        ),
        'reliability': DataValidation(
            type="list",
            formula1='"High,Medium,Low"',
            allow_blank=False
        ),
        'doc_requirement': DataValidation(
            type="list",
            formula1='"Mandatory,Optional,N/A"',
            allow_blank=False
        ),
        'format_type': DataValidation(
            type="list",
            formula1='"Electronic,Paper,Both"',
            allow_blank=False
        ),
        'disposal_method': DataValidation(
            type="list",
            formula1='"Secure Delete,Archive,Destruction"',
            allow_blank=False
        ),
        'deployment_type': DataValidation(
            type="list",
            formula1='"Cloud SaaS,On-Premises,Hybrid,Self-Hosted"',
            allow_blank=False
        ),
        'tool_capability': DataValidation(
            type="list",
            formula1=f'"{CHECK} Full,⚠ Partial,❌ No"',
            allow_blank=False
        ),
        'integration_type': DataValidation(
            type="list",
            formula1='"API,File Transfer,Manual,Webhook,Other"',
            allow_blank=False
        ),
        'integration_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Integrated,⚠ Partial,❌ Standalone"',
            allow_blank=False
        ),
        'license_status': DataValidation(
            type="list",
            formula1='"Active,Expiring Soon,Expired"',
            allow_blank=False
        ),
        'support_availability': DataValidation(
            type="list",
            formula1='"24/7,Business Hours,Community,None"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Change Request,Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other,N/A"',
            allow_blank=True
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⚠ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Final,⚠ Requires Remediation, Draft,❌ Re-assessment Required"',
            allow_blank=False
        ),
        'review_recommendation': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approve,⚠ Approve with Conditions,❌ Reject, Require Rework"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approved,⚠ Approved with Conditions,❌ Rejected"',
            allow_blank=False
        ),
        'regulatory_review': DataValidation(
            type="list",
            formula1='"Yes,No,To Be Determined"',
            allow_blank=False
        ),
    }

    # NOTE: Do NOT add validations here. They are added per-sheet via
    # finalize_validations() AFTER cells are assigned, to avoid empty
    # sqref entries that trigger Excel "Repaired" warnings.

    return validations


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Document YOUR organisation’s change management process in the Change Process Workflow sheet.', '2. Complete the Approval Authority Matrix with YOUR specific roles and thresholds.', '3. Fill in YOUR communication procedures and stakeholder notification methods.', '4. Document YOUR documentation requirements and record-keeping practices.', '5. Inventory YOUR change management tools (whatever platforms/systems you use).', '6. Review the Summary Dashboard for compliance metrics.', '7. Maintain the Evidence Register for audit traceability.', '8. Obtain final approval and sign-off.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_change_process_workflow(ws, styles):
    """Create Change Process Workflow sheet documenting the complete change lifecycle."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE MANAGEMENT PROCESS WORKFLOW"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document the complete change lifecycle in your organisation"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== PROCESS STAGE MAPPING ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "PROCESS STAGE MAPPING"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Stage", "Stage Name", "Description", "Process Owner", "Standard Duration", "Tool/System Used", "Status", "Evidence"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    stages = [
        (1, "Change Request Initiation", "How changes are requested"),
        (2, "Initial Assessment & Screening", "Validation and categorization"),
        (3, "Risk & Impact Assessment", "Evaluate change risks"),
        (4, "Change Categorization", "Standard/Normal/Emergency"),
        (5, "CAB Scheduling (if required)", "Schedule review meeting"),
        (6, "CAB Review & Discussion", "Formal review process"),
        (7, "Change Authorisation", "Approval decision"),
        (8, "Implementation Planning", "Schedule and resource allocation"),
        (9, "Stakeholder Communication", "Notification procedures"),
        (10, "Pre-Implementation Testing", "Testing in non-production"),
        (11, "Change Implementation", "Actual execution"),
        (12, "Implementation Verification", "Verify success"),
        (13, "Stakeholder Notification", "Completion notification"),
        (14, "Post-Implementation Review", "Lessons learned"),
        (15, "Change Record Closure", "Documentation completion"),
    ]

    row += 1
    for stage_num, stage_name, description in stages:
        ws[f"A{row}"] = stage_num
        ws[f"B{row}"] = stage_name
        ws[f"C{row}"] = description
        
        # Make editable cells yellow
        for col in [4, 5, 6, 8]:  # D=4, E=5, F=6, H=8
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Status dropdown
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        # Evidence dropdown
        validations['evidence_type'].add(ws.cell(row=row, column=8))

        row += 1

    # ==================== PROCESS DOCUMENTATION REQUIREMENTS ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "PROCESS DOCUMENTATION REQUIREMENTS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    doc_headers = ["Requirement", "Implemented", "Documentation Type", "Template Available", "Location", "Responsible Role"]
    for col_idx, header in enumerate(doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    doc_requirements = [
        "Change request form/template",
        "Risk assessment template",
        "Impact assessment template",
        "CAB agenda template",
        "Implementation plan template",
        "Rollback plan template",
        "PIR template",
        "Communication templates",
        "Change closure checklist",
    ]

    row += 1
    for requirement in doc_requirements:
        ws[f"A{row}"] = requirement
        
        # Editable cells
        for col in [3, 5, 6]:  # C=3, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['yes_no_partial'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]
        
        row += 1

    # ==================== WORKFLOW AUTOMATION ASSESSMENT ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "WORKFLOW AUTOMATION ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    auto_headers = ["Capability", "Automated", "Semi-Automated", "Manual", "Tool/System", "Notes"]
    for col_idx, header in enumerate(auto_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    automation_items = [
        "Change request submission",
        "Request routing/assignment",
        "Approval workflow",
        "Notification distribution",
        "Change calendar updates",
        "Status tracking",
        "Reporting/dashboards",
        "PIR reminders",
    ]

    row += 1
    for item in automation_items:
        ws[f"A{row}"] = item
        
        # Checkboxes (represented as Yes/No dropdowns in Excel)
        for col in [2, 3, 4]:  # B=2, C=3, D=4
            validations['yes_no'].add(ws.cell(row=row, column=col))
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws.cell(row=row, column=col).border = styles["border"]

        # Text fields
        for col in [5, 6]:  # E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5: APPROVAL AUTHORITY MATRIX SHEET
# ============================================================================

def create_approval_authority_matrix(ws, styles):
    """Create Approval Authority Matrix sheet defining approval authorities."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE APPROVAL AUTHORITY MATRIX"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Define approval authorities by change type and risk level"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== STANDARD CHANGES APPROVAL ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "STANDARD CHANGES APPROVAL"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    std_headers = ["Standard Change Type", "Pre-Approved", "Self-Service Allowed", "Approver (if required)", "Approval SLA", "Implemented", "Evidence"]
    for col_idx, header in enumerate(std_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Sample row with example data
    row += 1
    sample_data = {
        1: "Restart Web Server",
        2: "Yes",
        3: "Yes",
        4: "Web Admin Team Lead",
        5: "1 hour",
        6: "Implemented",
        7: "EV-STD-001",
    }
    for col_idx, value in sample_data.items():
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = styles["border"]

    # Add dropdowns to sample row
    validations['yes_no'].add(ws.cell(row=row, column=2))
    validations['yes_no_na'].add(ws.cell(row=row, column=3))
    validations['implementation_status'].add(ws.cell(row=row, column=6))
    validations['evidence_type'].add(ws.cell(row=row, column=7))

    # Empty data rows (50 rows for user data)
    row += 1
    for i in range(50):
        for col in range(1, 8):  # Columns A-G
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]
            cell.value = None  # Empty - users define their own standard changes

        # Add dropdowns to empty rows
        validations['yes_no'].add(ws.cell(row=row, column=2))
        validations['yes_no_na'].add(ws.cell(row=row, column=3))
        validations['implementation_status'].add(ws.cell(row=row, column=6))
        validations['evidence_type'].add(ws.cell(row=row, column=7))

        row += 1

    # ==================== NORMAL CHANGES APPROVAL MATRIX ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "NORMAL CHANGES APPROVAL MATRIX"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    normal_headers = ["Risk Level", "Impact Level", "Required Approvers", "CAB Review Required", "Approval Threshold", "Typical SLA", "Implemented"]
    for col_idx, header in enumerate(normal_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Define risk/impact combinations
    combinations = [
        ("Low", "Low"),
        ("Low", "Medium"),
        ("Low", "High"),
        ("Medium", "Low"),
        ("Medium", "Medium"),
        ("Medium", "High"),
        ("High", "Low"),
        ("High", "Medium"),
        ("High", "High"),
        ("Critical", "Any"),
    ]

    row += 1
    for risk, impact in combinations:
        ws[f"A{row}"] = risk
        ws[f"B{row}"] = impact
        
        # Editable cells
        for col in [3, 5, 6]:  # C=3, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        row += 1

    # ==================== EMERGENCY CHANGES APPROVAL ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGES APPROVAL"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    emerg_headers = ["Emergency Criteria", "E-CAB Members", "Minimum Approvers", "Approval Method", "Documented", "Evidence"]
    for col_idx, header in enumerate(emerg_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    emergency_criteria = [
        "System outage affecting critical services",
        "Security incident response",
        "Regulatory compliance deadline",
        "Data loss prevention",
        "[Other - specify]",
    ]

    row += 1
    for criteria in emergency_criteria:
        ws[f"A{row}"] = criteria

        # Editable cells
        for col in [2, 3, 6]:  # B=2, C=3, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['approval_method'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        # Evidence dropdown
        validations['evidence_type'].add(ws.cell(row=row, column=6))

        row += 1

    # ==================== APPROVAL PROCESS METRICS ====================
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "APPROVAL PROCESS METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Target", "Current", "Status"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        "Average approval time (Standard)",
        "Average approval time (Normal - Low Risk)",
        "Average approval time (Normal - High Risk)",
        "Average approval time (Emergency)",
        "Approval backlog (pending >SLA)",
        "Approval rejection rate",
    ]

    row += 1
    for metric in metrics:
        ws[f"A{row}"] = metric
        
        # Editable cells
        for col in [2, 3]:  # B=2, C=3
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Status - calculated (conditional formatting would be applied manually)
        ws[f"D{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 5.5: CAB OPERATIONS SHEET
# ============================================================================

def create_cab_operations(ws, styles):
    """Create CAB Operations sheet documenting Change Advisory Board composition and operation."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE ADVISORY BOARD (CAB) OPERATIONS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document CAB composition, meeting schedules, and operational procedures"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== CAB COMPOSITION ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CAB COMPOSITION & MEMBERSHIP"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    headers = ["Role", "Name", "Department", "Authority Level", "Mandatory/Optional", "Delegate", "Contact", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    cab_roles = [
        ("CAB Chair", "", "", "Final Decision", "Mandatory", "", "", ""),
        ("IT Security Representative", "", "", "Veto for Security", "Mandatory", "", "", ""),
        ("Infrastructure Lead", "", "", "Advisory", "Mandatory", "", "", ""),
        ("Application Lead", "", "", "Advisory", "Mandatory", "", "", ""),
        ("Business Representative", "", "", "Advisory", "As Needed", "", "", ""),
        ("Change Manager", "", "", "Facilitator", "Mandatory", "", "", ""),
    ]

    for role_data in cab_roles:
        for col_idx, value in enumerate(role_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 1:
                apply_style(cell, styles["input_cell"])
        row += 1

    row += 2

    # ==================== CAB MEETING SCHEDULE ====================
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CAB MEETING SCHEDULE & PROCEDURES"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    meeting_headers = ["Meeting Type", "Frequency", "Day/Time", "Duration", "Quorum Required", "Minutes Owner", "Agenda Deadline", "Status"]
    for col_idx, header in enumerate(meeting_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    meetings = [
        ("Regular CAB", "Weekly", "", "1 hour", "Chair + 3", "", "24 hours prior", ""),
        ("Emergency CAB (eCAB)", "As Needed", "", "30 mins", "Chair + 2", "", "Immediate", ""),
        ("CAB Review (Post-Implementation)", "Monthly", "", "2 hours", "Chair + 3", "", "48 hours prior", ""),
    ]

    for meeting_data in meetings:
        for col_idx, value in enumerate(meeting_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 1:
                apply_style(cell, styles["input_cell"])
        row += 1

    row += 2

    # ==================== CAB RESPONSIBILITIES ====================
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "CAB RESPONSIBILITIES & DECISION CRITERIA"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    resp_headers = ["Responsibility", "Description", "Owner", "Documented?", "Evidence Reference"]
    for col_idx, header in enumerate(resp_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    responsibilities = [
        ("Change Review", "Review all Normal and Major changes before implementation", "", "", ""),
        ("Risk Assessment", "Evaluate risk level and approve risk mitigation", "", "", ""),
        ("Schedule Coordination", "Ensure changes don't conflict with other activities", "", "", ""),
        ("Resource Approval", "Verify adequate resources for change implementation", "", "", ""),
        ("Post-Implementation Review", "Review PIR results and lessons learned", "", "", ""),
        ("Escalation Decisions", "Escalate changes requiring executive approval", "", "", ""),
    ]

    for resp_data in responsibilities:
        for col_idx, value in enumerate(resp_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles["input_cell"])
        # Evidence Reference dropdown (col 5 = E)
        validations['evidence_type'].add(ws.cell(row=row, column=5))
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 15

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 6: COMMUNICATION PROCEDURES SHEET
# ============================================================================

def create_communication_procedures(ws, styles):
    """Create Communication_Procedures sheet documenting stakeholder notifications."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE COMMUNICATION PROCEDURES"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Document stakeholder notification and communication methods"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== STAKEHOLDER IDENTIFICATION ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "STAKEHOLDER IDENTIFICATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    stakeholder_headers = ["Stakeholder Group", "Role/Description", "Notification Trigger", "Communication Method", "Responsible Party", "Template Available", "Status"]
    for col_idx, header in enumerate(stakeholder_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    stakeholder_groups = [
        "End Users",
        "Service Owners",
        "IT Operations Team",
        "Security Team",
        "Management",
        "CAB Members",
        "External Vendors",
        "[Other - specify]",
        "[Other - specify]",
        "[Other - specify]",
    ]

    row += 1
    for group in stakeholder_groups:
        ws[f"A{row}"] = group

        # Editable cells
        for col in [2, 5]:  # B=2, E=5
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['notification_trigger'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['communication_method'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        validations['yes_no'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        row += 1

    # ==================== COMMUNICATION TEMPLATES ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COMMUNICATION TEMPLATES"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    template_headers = ["Template Type", "Purpose", "Content Includes", "Approval Required", "Format", "Location", "Maintained"]
    for col_idx, header in enumerate(template_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    template_types = [
        ("Change Request Acknowledgment", "Confirm receipt", "Change ID, requestor, next steps"),
        ("Change Scheduled Notification", "Inform of approval & schedule", "Change details, date/time, impact"),
        ("Change Reminder (24h)", "Pre-implementation reminder", "Time, expected duration, contacts"),
        ("Change In Progress", "During implementation", "Status, progress, issues"),
        ("Change Completed Successfully", "Successful completion", "Completion time, verification"),
        ("Change Failed/Rolled Back", "Failure notification", "Issue, rollback status, next steps"),
        ("Change Postponed", "Schedule change", "Reason, new date, next steps"),
        ("Emergency Change Notice", "Urgent notification", "Reason, approval, impact"),
    ]

    row += 1
    for template_type, purpose, content in template_types:
        ws[f"A{row}"] = template_type
        ws[f"B{row}"] = purpose
        ws[f"C{row}"] = content

        # Editable cells
        for col in [6]:  # F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        # Format dropdown (Email/Portal/Both)
        format_dv = DataValidation(
            type="list",
            formula1='"Email,Portal,Both"',
            allow_blank=False
        )
        ws.add_data_validation(format_dv)
        format_dv.add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        
        row += 1

    # ==================== COMMUNICATION CHANNELS ASSESSMENT ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMMUNICATION CHANNELS ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    channel_headers = ["Channel", "Available", "Primary Use", "Audience Reach", "Reliability", "Documented"]
    for col_idx, header in enumerate(channel_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    channels = [
        "Email",
        "Internal Portal/Intranet",
        "Change Management System",
        "Instant Messaging (Teams/Slack/etc)",
        "SMS/Text Messages",
        "Status Dashboard",
        "Scheduled Announcements",
    ]

    row += 1
    for channel in channels:
        ws[f"A{row}"] = channel

        # Editable cells
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        # Dropdowns
        validations['yes_no'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['audience_reach'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].border = styles["border"]

        validations['reliability'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].border = styles["border"]
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 7: DOCUMENTATION REQUIREMENTS SHEET
# ============================================================================

def create_documentation_requirements(ws, styles):
    """Create Documentation_Requirements sheet defining record-keeping requirements."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE DOCUMENTATION REQUIREMENTS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "Define record-keeping requirements by change type"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== STANDARD CHANGE DOCUMENTATION ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "STANDARD CHANGE DOCUMENTATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    std_doc_headers = ["Documentation Item", "Required", "Format", "Retention Period", "Storage Location", "Responsible Role", "Compliant"]
    for col_idx, header in enumerate(std_doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    std_doc_items = [
        "Change request (basic)",
        "Requestor identification",
        "Change description",
        "Service/system affected",
        "Implementation date/time",
        "Implementation result",
    ]

    row += 1
    for item in std_doc_items:
        ws[f"A{row}"] = item

        # Editable cells
        for col in [4, 5, 6]:  # D=4, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['doc_requirement'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['format_type'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        row += 1

    # ==================== NORMAL CHANGE DOCUMENTATION ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "NORMAL CHANGE DOCUMENTATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(std_doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    normal_doc_items = [
        "Complete change request",
        "Risk assessment",
        "Impact assessment",
        "Business justification",
        "Technical implementation plan",
        "Test plan/results",
        "Rollback plan",
        "CAB review record",
        "Approval record (all approvers)",
        "Communication records",
        "Implementation log/notes",
        "Verification evidence",
        "Post-Implementation Review",
    ]

    row += 1
    for item in normal_doc_items:
        ws[f"A{row}"] = item

        # Editable cells
        for col in [4, 5, 6]:  # D=4, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['doc_requirement'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['format_type'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]

        row += 1

    # Additional ISO 27002:2022 normal change documentation items (F2F2F2 reference rows)
    extra_normal_items = [
        "Downstream documentation updated — policies, procedures, and operating manuals reviewed and updated where affected by this change",
        "BCM/ICT continuity plan reviewed — business continuity and disaster recovery plans updated if system criticality or architecture has changed",
    ]
    for item in extra_normal_items:
        ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, color="003366")
        row += 1

    # ==================== EMERGENCY CHANGE DOCUMENTATION ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGE DOCUMENTATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    emerg_doc_headers = ["Documentation Item", "Required", "Special Requirements", "Retention Period", "Storage Location", "Responsible Role", "Compliant"]
    for col_idx, header in enumerate(emerg_doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    emerg_doc_items = [
        "Emergency change justification",
        "E-CAB approval record",
        "Emergency criteria met",
        "Incident ticket reference",
        "Risk accepted by",
        "Implementation record",
        "Mandatory PIR",
        "Retrospective CAB review",
    ]

    row += 1
    for item in emerg_doc_items:
        ws[f"A{row}"] = item

        # Editable cells
        for col in [3, 4, 5, 6]:  # C=3, D=4, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['doc_requirement'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].border = styles["border"]
        
        row += 1

    # ==================== RECORD RETENTION & DISPOSAL ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "RECORD RETENTION & DISPOSAL"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    retention_headers = ["Record Type", "Retention Period", "Disposal Method", "Regulatory Requirement", "Implemented", "Evidence"]
    for col_idx, header in enumerate(retention_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    record_types = [
        "Standard changes",
        "Normal changes (successful)",
        "Failed changes",
        "Emergency changes",
        "CAB meeting minutes",
        "Approval records",
    ]

    row += 1
    for record_type in record_types:
        ws[f"A{row}"] = record_type

        # Editable cells
        for col in [2, 4, 6]:  # B=2, D=4, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdowns
        validations['disposal_method'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].border = styles["border"]

        validations['implementation_status'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].border = styles["border"]

        # Evidence dropdown (col 6 = F)
        validations['evidence_type'].add(ws.cell(row=row, column=6))

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8: CHANGE MANAGEMENT TOOLS SHEET
# ============================================================================

def create_change_management_tools(ws, styles):
    """Create Change_Management_Tools sheet inventorying all supporting tools/systems."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "CHANGE MANAGEMENT TOOLS & SYSTEMS INVENTORY"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = "Document all platforms and tools supporting change management"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== PRIMARY CHANGE MANAGEMENT SYSTEM ====================
    row = 4
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "PRIMARY CHANGE MANAGEMENT SYSTEM"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    ws[f"A{row}"] = "Attribute"
    ws[f"B{row}"] = "Value"
    ws[f"C{row}"] = "Evidence Location"
    for col in [1, 2, 3]:  # A=1, B=2, C=3
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    primary_system_fields = [
        ("SYSTEM IDENTIFICATION", None),
        ("Tool/Platform Name", "text"),
        ("Vendor/Provider", "text"),
        ("Version/Release", "text"),
        ("Deployment Type", "deployment_type"),
        ("Primary Use", "text"),
        ("", None),
        ("CAPABILITIES", None),
        ("Change request submission", "tool_capability"),
        ("Workflow automation", "tool_capability"),
        ("Approval routing", "tool_capability"),
        ("Risk/Impact assessment forms", "tool_capability"),
        ("Change calendar", "tool_capability"),
        ("Conflict detection", "tool_capability"),
        ("Notification engine", "tool_capability"),
        ("Reporting/dashboards", "tool_capability"),
        ("Integration with monitoring", "tool_capability"),
        ("Integration with CMDB", "tool_capability"),
        ("API availability", "tool_capability"),
        ("", None),
        ("ADMINISTRATION", None),
        ("System administrators", "text"),
        ("User training provided", "yes_no"),
        ("Documentation available", "yes_no"),
        ("Support availability", "support_availability"),
        ("License status", "license_status"),
        ("License expiration date", "date"),
    ]

    row += 1
    for field_name, field_type in primary_system_fields:
        ws[f"A{row}"] = field_name
        
        if field_name and field_name.isupper():
            # Section header
            ws[f"A{row}"].font = Font(bold=True, italic=True)
        elif field_type:
            # Editable value cell
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = styles["border"]
            ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"C{row}"].border = styles["border"]
            
            # Apply validation if specified
            if field_type == "date":
                ws[f"B{row}"].number_format = 'DD.MM.YYYY'
            elif field_type != "text" and field_type in validations:
                validations[field_type].add(ws[f"B{row}"])
        
        row += 1

    # ==================== SUPPORTING TOOLS & SYSTEMS ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "SUPPORTING TOOLS & SYSTEMS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    support_headers = ["Tool/System", "Purpose", "Integration with Primary", "Status", "User Count", "Evidence"]
    for col_idx, header in enumerate(support_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    row += 1
    # 18 rows for supporting tools
    for i in range(18):
        for col in [1, 2, 5, 6]:  # A=1, B=2, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        validations['integration_status'].add(ws.cell(row=row, column=3))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=3).border = styles["border"]

        validations['implementation_status'].add(ws.cell(row=row, column=4))
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=4).border = styles["border"]

        # Evidence dropdown (col 6 = F)
        validations['evidence_type'].add(ws.cell(row=row, column=6))

        row += 1

    # ==================== INTEGRATION ARCHITECTURE ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "INTEGRATION ARCHITECTURE"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    integ_headers = ["Integration Point", "Source System", "Target System", "Integration Type", "Automated", "Data Exchanged", "Status"]
    for col_idx, header in enumerate(integ_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    row += 1
    # 12 rows for integrations
    for i in range(12):
        for col in [1, 2, 3, 6]:  # A=1, B=2, C=3, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        validations['integration_type'].add(ws.cell(row=row, column=4))
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=4).border = styles["border"]

        validations['yes_no_partial'].add(ws.cell(row=row, column=5))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=5).border = styles["border"]

        validations['implementation_status'].add(ws.cell(row=row, column=7))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=row, column=7).border = styles["border"]
        
        row += 1

    # ==================== TOOL ACCESS & SECURITY ====================
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "TOOL ACCESS & SECURITY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    security_headers = ["Security Control", "Implemented", "Details", "Evidence"]
    for col_idx, header in enumerate(security_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    security_controls = [
        "Multi-factor authentication",
        "Role-based access control",
        "Audit logging enabled",
        "Data encryption (at rest)",
        "Data encryption (in transit)",
        "Backup/recovery tested",
        "Disaster recovery plan",
    ]

    row += 1
    for control in security_controls:
        ws[f"A{row}"] = control

        # Editable cells
        for col in [3, 4]:  # C=3, D=4
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = styles["border"]

        # Dropdown
        validations['yes_no_partial'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = styles["border"]

        # Evidence dropdown (col 4 = D)
        validations['evidence_type'].add(ws.cell(row=row, column=4))

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 8.5: METRICS & KPIS SHEET
# ============================================================================

def create_metrics_kpis(ws, styles):
    """Create Metrics KPIs sheet documenting tracked change management metrics."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE MANAGEMENT METRICS AND KPIS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document tracked metrics, targets, and reporting frequency"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== KEY METRICS ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY CHANGE MANAGEMENT METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    headers = ["Metric Name", "Description", "Target", "Current Value", "Frequency", "Data Source", "Owner", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    metrics = [
        ("Change Success Rate", "% of changes implemented without incident", ">95%", "", "Monthly", "", "", ""),
        ("Emergency Change Rate", "% of changes classified as emergency", "<10%", "", "Monthly", "", "", ""),
        ("Change Backlog", "Number of approved changes awaiting implementation", "<20", "", "Weekly", "", "", ""),
        ("Mean Time to Implement", "Average time from approval to implementation", "<5 days", "", "Monthly", "", "", ""),
        ("Failed Change Rate", "% of changes requiring rollback", "<5%", "", "Monthly", "", "", ""),
        ("CAB Approval Time", "Average time to CAB decision", "<48 hours", "", "Monthly", "", "", ""),
        ("Post-Implementation Review Rate", "% of changes with completed PIR", ">90%", "", "Monthly", "", "", ""),
        ("Change-Related Incidents", "Incidents caused by changes", "<5/month", "", "Monthly", "", "", ""),
    ]

    for metric_data in metrics:
        for col_idx, value in enumerate(metric_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 3:
                apply_style(cell, styles["input_cell"])
        row += 1

    row += 2

    # ==================== REPORTING SCHEDULE ====================
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "METRICS REPORTING SCHEDULE"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    report_headers = ["Report Name", "Audience", "Frequency", "Format", "Delivery Method", "Owner", "Status"]
    for col_idx, header in enumerate(report_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    reports = [
        ("Weekly Change Summary", "IT Management", "Weekly", "Dashboard", "", "", ""),
        ("Monthly Change Report", "CISO/Executive", "Monthly", "PDF Report", "", "", ""),
        ("Quarterly Trend Analysis", "Board/Audit", "Quarterly", "Presentation", "", "", ""),
        ("Annual Change Review", "Executive/Audit", "Annual", "Full Report", "", "", ""),
    ]

    for report_data in reports:
        for col_idx, value in enumerate(report_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles["input_cell"])
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 15

    ws.freeze_panes = "A6"
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)

# ============================================================================
# SECTION 9: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with standard compliance table and metrics."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE PROCESS ASSESSMENT — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.8.32: Change Management"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty (for standardisation)

    # --- TABLE 1: COMPLIANCE OVERVIEW ---
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # TABLE 1 assessment areas with formulas
    # Pattern based on A.8.33-34 (area_configs with dynamic formula generation)
    # Based on ISO 27001:2022 A.8.32 requirements analysis

    area_configs = [
        # (Sheet Name, Status Column, [Compliant, Partial, Non-Compliant], Row Start, Row End)
        ('Change Process Workflow', 'G', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 7, 56),
        # Approval Authority Matrix - 3 sections with formulas (extended to cover all data to row 89)
        ('Approval Authority Matrix', 'F', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 7, 89),
        ('Approval Authority Matrix', 'G', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 61, 89),
        ('Approval Authority Matrix', 'E', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 75, 89),
        ('Communication', 'G', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 6, 55),
        ('Documentation Records', 'G', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 6, 55),
        # Change Management Tools - 2 sections with formulas
        ('Tool Capabilities', 'D', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 36, 81),
        ('Tool Capabilities', 'G', ['✅ Implemented', '⚠ Partial', '❌ Not Implemented'], 57, 81),
    ]

    assessment_areas = [
        'Change Process Workflow',
        'Approval Authority Matrix - Standard Changes',
        'Approval Authority Matrix - Normal Changes',
        'Approval Authority Matrix - Emergency Changes',
        'Communication',
        'Documentation Records',
        'Change Management Tools - Supporting Tools',
        'Change Management Tools - Integration Architecture',
    ]

    r = 6
    for i, area_name in enumerate(assessment_areas):
        # Column A: Area name
        ws[f"A{r}"] = area_name
        ws[f"A{r}"].font = Font(name="Calibri", size=10)
        ws[f"A{r}"].border = border

        if area_configs[i] is None:
            # MANUAL ENTRY (no compliance column)
            for col in "BCDEF":
                cell = ws[f"{col}{r}"]
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws[f"G{r}"].border = border
            ws[f"G{r}"].alignment = Alignment(horizontal="center", vertical="center")
        else:
            # FORMULA-BASED
            sheet_name, status_col, status_values, row_start, row_end = area_configs[i]

            # B: Total
            ws[f"B{r}"] = f"=COUNTA('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end})"
            ws[f"B{r}"].border = border
            ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # C: Compliant
            ws[f"C{r}"] = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[0]}")'
            ws[f"C{r}"].border = border
            ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # D: Partial
            ws[f"D{r}"] = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[1]}")'
            ws[f"D{r}"].border = border
            ws[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # E: Non-Compliant (includes "Planned")
            ws[f"E{r}"] = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"{status_values[2]}")+COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"Planned")'
            ws[f"E{r}"].border = border
            ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # F: N/A
            ws[f"F{r}"] = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"N/A")'
            ws[f"F{r}"].border = border
            ws[f"F{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # G: Compliance % (will be added by TOTAL row loop below)
            ws[f"G{r}"].border = border
            ws[f"G{r}"].alignment = Alignment(horizontal="center", vertical="center")

        r += 1

    # Add Compliance % formulas for each assessment area row
    for row_idx in range(6, 14):  # rows 6-13 (8 assessment areas)
        ws[f"G{row_idx}"] = f'=IF((B{row_idx}-F{row_idx})=0,0,C{row_idx}/(B{row_idx}-F{row_idx}))'
        ws[f"G{row_idx}"].number_format = "0.0%"
        ws[f"G{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")

    # TOTAL row with SUM formulas
    total_row = 14  # After 8 assessment areas (rows 6-13)
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{total_row}"].border = border
    for col_letter in "BCDEF":
        cell = ws[f"{col_letter}{total_row}"]
        cell.value = f"=SUM({col_letter}6:{col_letter}13)"
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"G{total_row}"] = f'=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))'
    ws[f"G{total_row}"].number_format = "0.0%"
    ws[f"G{total_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"G{total_row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"G{total_row}"].border = border
    ws[f"G{total_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # --- TABLE 2: KEY METRICS ---
    met_row = total_row + 2
    ws.merge_cells(f"A{met_row}:G{met_row}")
    ws[f"A{met_row}"] = "TABLE 2: KEY METRICS"
    ws[f"A{met_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{met_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{met_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=met_row, column=col).border = border

    # TABLE 2 column headers (D9D9D9 grey)
    t2_hdr_row = met_row + 1
    t2_headers = ["Metric", "Value", "", "", "", "", ""]
    for col_idx, hdr in enumerate(t2_headers, 1):
        hcell = ws.cell(row=t2_hdr_row, column=col_idx, value=hdr)
        hcell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        hcell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        hcell.border = border
        hcell.alignment = Alignment(horizontal="center", vertical="center")

    # Metrics with formulas (where available) or placeholders
    metrics = [
        ("Change Success Rate", "='Metrics KPIs'!D6"),
        ("Emergency Change Rate", "='Metrics KPIs'!D7"),
        ("Average Change Lead Time", "[enter days]"),
        ("Failed Change Rate", "='Metrics KPIs'!D10"),
    ]
    r = t2_hdr_row
    for metric, value in metrics:
        r += 1
        ws[f"A{r}"] = metric
        ws[f"A{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"A{r}"].border = border
        ws[f"B{r}"] = value
        ws[f"B{r}"].font = Font(name="Calibri", size=10, color="000000")
        ws[f"B{r}"].border = border
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = border

    # TABLE 2 buffer rows (2 empty white rows)
    for _ in range(2):
        r += 1
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border

    # --- TABLE 3: CRITICAL FINDINGS ---
    crit_start = r + 2
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{crit_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=crit_start, column=col).border = border

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col_idx, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Critical findings based on change management control requirements
    findings = [
        ("Change Process", "Changes without approval", '=COUNTIF(\'Change Process Workflow\'!G7:G56,"\u274c Not Implemented")', "Critical", "Immediate"),
        ("Approval Authority", "Emergency changes without approval", '=COUNTIF(\'Approval Authority Matrix\'!E75:E89,"\u274c Not Implemented")', "Critical", "Immediate"),
        ("Communication", "Changes without stakeholder notification", '=COUNTIF(\'Communication\'!G6:G55,"\u274c Not Implemented")', "Critical", "Immediate"),
        ("Documentation", "Changes without documentation", '=COUNTIF(\'Documentation Records\'!G6:G55,"\u274c Not Implemented")', "Critical", "Immediate"),
        ("Tool Capabilities", "Critical tool capability gaps", '=COUNTIF(\'Tool Capabilities\'!D36:D81,"\u274c Not Implemented")', "Critical", "Immediate"),
        ("Change Process", "High-risk changes with partial implementation", '=COUNTIF(\'Change Process Workflow\'!G7:G56,"\u26a0 Partial")', "High", "Urgent"),
        ("Approval Authority", "Standard changes with approval gaps", '=COUNTIF(\'Approval Authority Matrix\'!F7:F89,"\u26a0 Partial")', "High", "Urgent"),
        ("Approval Authority", "Normal changes with approval gaps", '=COUNTIF(\'Approval Authority Matrix\'!G61:G89,"\u26a0 Partial")', "High", "Urgent"),
        ("Communication", "Partial communication implementation", '=COUNTIF(\'Communication\'!G6:G55,"\u26a0 Partial")', "Medium", "Plan"),
        ("Tool Capabilities", "Tool workflow gaps", '=COUNTIF(\'Tool Capabilities\'!G57:G81,"\u26a0 Partial")', "Medium", "Plan"),
    ]

    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row = crit_start + 1
    for cat, finding, formula, severity, action in findings:
        row += 1
        ws.cell(row=row, column=1, value=cat).border = border
        ws.cell(row=row, column=1).font = Font(color="000000")
        ws.cell(row=row, column=1).fill = ffffcc_fill
        ws.cell(row=row, column=2, value=finding).border = border
        ws.cell(row=row, column=2).font = Font(color="000000")
        ws.cell(row=row, column=2).fill = ffffcc_fill
        cell_count = ws.cell(row=row, column=3)
        cell_count.value = formula
        cell_count.border = border
        cell_count.alignment = Alignment(horizontal="center")
        cell_count.font = Font(color="000000")
        cell_count.fill = ffffcc_fill
        ws.cell(row=row, column=4, value=severity).border = border
        ws.cell(row=row, column=4).font = Font(color="000000")
        ws.cell(row=row, column=4).fill = ffffcc_fill
        ws.cell(row=row, column=5, value=action).border = border
        ws.cell(row=row, column=5).font = Font(color="000000")
        ws.cell(row=row, column=5).fill = ffffcc_fill
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = ffffcc_fill
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).fill = ffffcc_fill

    # TABLE 3 buffer rows (2 empty FFFFCC rows)
    for _ in range(2):
        row += 1
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border

    # Column widths (19 columns)
    gap_widths = [
        ("A", 15), ("B", 20), ("C", 40), ("D", 30), ("E", 30),
        ("F", 10), ("G", 12), ("H", 12), ("I", 12), ("J", 35),
        ("K", 20), ("L", 15), ("M", 15), ("N", 12), ("O", 15),
        ("P", 25), ("Q", 25), ("R", 30), ("S", 20)
    ]
    for col, w in gap_widths:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create standard 8-column Evidence Register with 100 rows."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Evidence type dropdown
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ev_type_dv.error = "Please select a valid evidence type"
    ev_type_dv.errorTitle = "Invalid Evidence Type"
    ws.add_data_validation(ev_type_dv)

    # Verification status dropdown
    ver_status_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Verified,{WARNING} Pending,{XMARK} Not Verified,N/A"',
        allow_blank=True,
    )
    ver_status_dv.error = "Please select a valid status"
    ver_status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(ver_status_dv)

    # Sample row (row 5) with example data (grey fill per Option B standard)
    sample_data = {
        1: "EV-001",
        2: "Change Process Assessment",
        3: "Policy Document",
        4: "Change Management Policy v1.0",
        5: "\\\\fileserver\\policies\\change_mgmt_v1.0.pdf",
        6: "15.01.2026",
        7: "Change Manager",
        8: f"{CHECK} Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Empty data rows (rows 6-105) - 100 empty rows for user data (MAX-002 standard)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            cell.value = None  # Empty - users choose their own evidence IDs
        # Dropdowns
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths
    for col, w in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create standard Approval Sign-Off with 3-section approval workflow."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    # Apply borders to all merged cells in header row
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}1"].border = border

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    # Apply borders to all cells in merged range
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}2"].border = border

    # --- ASSESSMENT SUMMARY ---
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    # Assessment Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.32.1 \u2014 Change Process Assessment", False),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G14", False),  # TABLE 1 TOTAL row
        ("Assessment Period:", "", True),
        ("Assessed By:", "", True),
        ("Assessment Status:", "", "dropdown"),
    ]
    row = 5
    for label, value, editable in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if editable == "dropdown":
            status_dv.add(ws[f"B{row}"])
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        elif editable:
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].border = border
        else:
            # Non-editable fields still need borders
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].border = border
        row += 1

    # --- Helper for approver sections ---
    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{start_row}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        # Apply borders to all merged cells in section title
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{start_row}"].border = border
        r = start_row + 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = field
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = border
            ws.merge_cells(f"B{r}:E{r}")
            # Apply borders and fill to all merged cells B-E
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{r}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{r}"].border = border
            r += 1
        return r

    # COMPLETED BY
    row += 1
    row = _approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")

    # REVIEWED BY
    row += 1
    row = _approver_section(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")

    # APPROVED BY (CISO)
    row += 1
    row = _approver_section(row, "APPROVED BY \u2014 CISO", "003366")

    # --- FINAL DECISION ---
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    # Apply borders and fill to all merged cells B-E
    for col in ["B", "C", "D", "E"]:
        ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"{col}{row}"].border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # --- NEXT REVIEW DETAILS ---
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to all merged cells in section title
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = border

    for field in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        row += 1
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        # Apply borders and fill to all merged cells B-E
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].border = border

    # Column widths
    for col, w in [("A", 32), ("B", 25), ("C", 20), ("D", 20), ("E", 20)]:
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "The first principle is that you must not fool yourself—
    and you are the easiest person to fool." - Richard Feynman
    
    This script creates evidence-based assessment tools for change management,
    not checkbox compliance theater.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.32.1 - Change Process Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.32: Change Management")
    logger.info("=" * 78)
    logger.info("\n Systems Engineering Approach: Evidence-Based Compliance")
    logger.info(f" Technology-Agnostic: Works with ANY change management approach")
    logger.info(f" Audit-Ready: Comprehensive evidence collection")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    logger.info("{CHECK} Workbook created with 11 sheets")

    # Create all sheets (per IMP specification - 11 sheets)
    logger.info("\n[Phase 2] Generating assessment sheets...")

    logger.info("  [1/11] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✅ Instructions complete")

    logger.info("  [2/11] Creating Change Process Workflow...")
    create_change_process_workflow(wb["Change Process Workflow"], styles)
    logger.info("  ✅ Process workflow complete (15 stages, automation assessment)")

    logger.info("  [3/11] Creating Approval Authority Matrix...")
    create_approval_authority_matrix(wb["Approval Authority Matrix"], styles)
    logger.info("  ✅ Approval matrix complete (Standard/Normal/Emergency changes)")

    logger.info("  [4/11] Creating CAB Operations...")
    create_cab_operations(wb["CAB Operations"], styles)
    logger.info("  ✅ CAB Operations complete (composition, meetings, responsibilities)")

    logger.info("  [5/11] Creating Communication...")
    create_communication_procedures(wb["Communication"], styles)
    logger.info("  ✅ Communication complete (stakeholder notifications)")

    logger.info("  [6/11] Creating Documentation Records...")
    create_documentation_requirements(wb["Documentation Records"], styles)
    logger.info("  ✅ Documentation records complete (record retention)")

    logger.info("  [7/11] Creating Tool Capabilities...")
    create_change_management_tools(wb["Tool Capabilities"], styles)
    logger.info("  ✅ Tool capabilities complete (primary system + integrations)")

    logger.info("  [8/11] Creating Metrics KPIs...")
    create_metrics_kpis(wb["Metrics KPIs"], styles)
    logger.info("  ✅ Metrics KPIs complete (tracked metrics and reporting)")

    logger.info("  [9/11] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    logger.info("  ✅ Evidence register complete (100 evidence rows)")

    logger.info("  [10/11] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    logger.info("  ✅ Dashboard complete (compliance metrics, audit readiness)")

    logger.info("  [11/11] Creating Approval Sign-Off...")
    create_approval_sheet(wb["Approval Sign-Off"], styles)
    logger.info("  ✅ Approval workflow complete (3-level sign-off)")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.32.1_Change_Process_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"{XMARK} ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info(" WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n Assessment Sheets (11 per IMP specification):")
    logger.info("  • Instructions & Legend (usage guidance)")
    logger.info("  • Change Process Workflow (15-stage lifecycle)")
    logger.info("  • Approval Authority Matrix (Standard/Normal/Emergency)")
    logger.info("  • CAB Operations (composition, meetings, responsibilities)")
    logger.info("  • Communication (stakeholder notifications)")
    logger.info("  • Documentation Records (record-keeping)")
    logger.info("  • Tool Capabilities (tool inventory + integrations)")
    logger.info("  • Metrics KPIs (tracked metrics and reporting)")
    logger.info("\n Analysis & Governance:")
    logger.info("  • Evidence Register (100 evidence entries)")
    logger.info("  • Summary Dashboard (compliance metrics, audit readiness)")
    logger.info("  • Approval Sign-Off (3-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info(" ASSESSMENT CAPABILITIES:")
    logger.info("  • 15 process stages documented")
    logger.info("  • 20 policy requirements mapped")
    logger.info("  • Standard/Normal/Emergency change approval paths")
    logger.info("  • Stakeholder communication procedures")
    logger.info("  • Documentation & retention requirements")
    logger.info("  • Tool capability assessment")
    logger.info("  • 100 evidence documentation entries")
    logger.info("  • Automated compliance calculations")
    logger.info("\n" + "─" * 78)
    logger.info(f" KEY FEATURES:")
    logger.info("  ✅ Technology-agnostic (works with ANY change management tool)")
    logger.info("  ✅ Comprehensive evidence collection")
    logger.info("  ✅ Automated compliance calculations")
    logger.info("  ✅ Audit readiness assessment")
    logger.info("  ✅ Multi-level approval workflow")
    logger.info("  ✅ Quarterly review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(f" NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Review Instructions & Legend sheet first")
    logger.info("  3. Document YOUR organisation's change process")
    logger.info("  4. Complete approval authority definitions")
    logger.info("  5. Define communication procedures")
    logger.info("  6. Document record-keeping requirements")
    logger.info("  7. Inventory YOUR change management tools")
    logger.info("  8. Review Summary Dashboard for compliance status")
    logger.info("  9. Document evidence in Evidence Register")
    logger.info("  10. Obtain final approval via Approval Sign-Off")
    logger.info("\n PRO TIP:")
    logger.info("  This workbook is technology-independent. Whether you use ServiceNow,")
    logger.info("  Jira, custom tools, or spreadsheets - this framework assesses YOUR")
    logger.info("  change management PROCESS, not your tool brands. That's the difference")
    logger.info("  between compliance theater and Systems Engineering.")
    logger.info("\n" + "=" * 78)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info("\n This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
