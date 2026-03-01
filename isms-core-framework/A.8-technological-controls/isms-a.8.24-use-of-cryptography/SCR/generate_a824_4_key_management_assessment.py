#!/usr/bin/env python3
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.24.4 - Key Management Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography
Assessment Domain 4 of 4: Cryptographic Key Management Lifecycle

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific key management infrastructure, cryptographic
standards, and assessment requirements.

Key customization areas:
1. Key management systems (match your actual KMS/HSM/PKI infrastructure)
2. Key lifecycle requirements (adapt to your operational procedures)
3. Key rotation schedules (specific to your risk profile and compliance needs)
4. Access control requirements (based on your authorisation model)
5. Compliance thresholds (aligned with your regulatory requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Use of Cryptography Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
cryptographic key management lifecycle controls across all key types and
management systems.

**Purpose:**
Enables systematic assessment of key management practices against ISO 27001:2022
Control A.8.24 requirements, supporting evidence-based validation of cryptographic
key lifecycle management from generation through destruction.

**Assessment Scope:**
- Key generation (algorithms, entropy sources, randomness)
- Key storage (KMS, HSM, software keystores, TPM)
- Key distribution and escrow
- Key rotation and renewal schedules
- Key backup and recovery procedures
- Key revocation and suspension
- Key archival and destruction
- Access control and separation of duties
- Audit logging and monitoring
- Certificate authority operations (if applicable)
- Hardware security module (HSM) usage and management
- Key ceremony procedures for critical keys
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and key management standards
2. Key Inventory - Comprehensive inventory of all cryptographic keys
3. Key Generation - Key creation processes and entropy assessment
4. Key Storage - KMS, HSM, and secure storage evaluation
5. Key Distribution - Key distribution and escrow procedures
6. Key Rotation - Rotation schedules and compliance tracking
7. Key Backup & Recovery - Backup procedures and recovery testing
8. Key Revocation - Revocation and suspension processes
9. Key Destruction - Secure key disposal and destruction
10. Access Control - Key access authorisation and separation of duties
11. Audit & Monitoring - Key usage logging and security monitoring
12. CA Operations - Certificate authority operations (if applicable)
13. HSM Management - Hardware security module configuration and operations
14. Gap Analysis - Key management deficiencies and remediation requirements
15. Evidence Register - Audit evidence tracking and documentation
16. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with key management standard dropdown lists
- Conditional formatting for lifecycle compliance status
- Automated gap identification for overdue rotation or missing controls
- Protected formulas with unprotected input cells
- Key inventory to lifecycle control mapping
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with key management systems

**Integration:**
This assessment feeds into the A.8.24.5 Compliance Dashboard, which
consolidates data from all four cryptographic assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a824_4_key_management_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a824_4_key_management_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a824_4_key_management_assessment.py --date 20250115

Output:
    File: ISMS_A_8_24_4_Key_Management_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize key management standards to match your risk profile
    2. Inventory all cryptographic keys across infrastructure
    3. Complete lifecycle assessments for each key type
    4. Validate key rotation schedules and compliance
    5. Review key storage security (KMS/HSM configuration)
    6. Assess access controls and separation of duties
    7. Conduct gap analysis for lifecycle deficiencies
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (KMS logs, rotation reports)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.24.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.24
Assessment Domain:    4 of 4 (Cryptographic Key Management Lifecycle)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.24: Use of Cryptography Policy (Governance)
    - ISMS-IMP-A.8.24.1: Data Transmission Cryptography Assessment (Domain 1)
    - ISMS-IMP-A.8.24.2: Data Storage Cryptography Assessment (Domain 2)
    - ISMS-IMP-A.8.24.3: Authentication Cryptography Assessment (Domain 3)
    - ISMS-IMP-A.8.24.4: Key Management Implementation Guide
    - ISMS-IMP-A.8.24.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.24.4 specification
    - Supports comprehensive key management lifecycle evaluation
    - Integrated with A.8.24.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Cryptographic Standards:**
Key management standards evolve based on cryptographic advances and threats.
Review industry standards (NIST SP 800-57, ISO/IEC 19790, FIPS 140-3) annually
and update assessment criteria accordingly. Ensure compliance with key length
recommendations and rotation schedules.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of key lifecycle controls, particularly
rotation schedules, access controls, and destruction procedures.

**Data Protection:**
Assessment workbooks contain HIGHLY sensitive infrastructure details including:
- Cryptographic key inventory and metadata
- Key management system architecture
- Key storage locations and access controls
- Key lifecycle deficiencies (critical security gaps)

Handle with MAXIMUM SECURITY in accordance with your organisation's highest
data classification level. Consider encrypting the assessment workbook itself.

**Maintenance:**
Review and update assessment:
- Monthly: Check key rotation compliance and overdue rotations
- Quarterly: Validate access controls and separation of duties
- Semi-annually: Review key management procedures and update if needed
- Annually: Complete reassessment of all key management infrastructure
- Ad-hoc: When cryptographic incidents or key compromises occur

**Quality Assurance:**
Have cryptography specialists and key management system administrators validate
assessments before using results for compliance reporting or remediation
decisions. Key management errors can have catastrophic security consequences.

**Regulatory Alignment:**
Ensure key management practices align with applicable requirements:
- Payment processing: PCI DSS v4.0.1 key management requirements (especially 3.5-3.7)
- Healthcare: HIPAA key management for encryption
- Privacy: GDPR encryption key protection requirements
- Finance: Regional banking cryptographic key management standards
- Government: Jurisdiction-specific key management mandates (e.g., FIPS 140-3)

Customize assessment criteria to include regulatory-specific requirements.

**Business Continuity:**
Key loss or unavailability can result in permanent data loss. Ensure key
backup and recovery procedures are thoroughly tested and documented. Balance
availability requirements with security controls (key escrow, split knowledge).

**Separation of Duties:**
Key management access should follow principle of least privilege with
appropriate separation of duties. No single individual should control entire
key lifecycle. Document and enforce role-based access controls.

**HSM Considerations:**
If using Hardware Security Modules (HSMs), ensure proper configuration,
monitoring, and maintenance. HSMs should be FIPS 140-2 Level 2+ or FIPS 140-3
compliant for production cryptographic operations.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.24.4"
WORKBOOK_NAME = "Key Management Assessment"
CONTROL_ID = "A.8.24"
CONTROL_NAME = "Use of Cryptography"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches markdown specification
    sheets = [
        "Instructions & Legend",
        "1. Key Generation",
        "2. Key Storage",
        "3. Key Rotation",
        "4. Key Backup & Recovery",
        "5. Certificate Management",
        "6. Key Destruction",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "exception_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
    }
    return styles



_STYLES = setup_styles()
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 2: COLUMN DEFINITIONS (BASE + EXTENSIONS)
# ============================================================================

def get_base_key_mgmt_columns():
    """
    Return base 17 columns (A-Q) standard for ALL key management assessment sheets.
    
    Returns:
        dict: {column_name: width}
    """
    return {
        "System/Service": 30,
        "Key Type": 20,
        "Algorithm": 18,
        "Key Length (bits)": 16,
        "Storage Method": 22,
        "Data Classification": 18,
        "Lifecycle Stage": 18,
        "Status": 15,
        "Evidence Location": 28,
        "Gap Description": 30,
        "Remediation Needed": 14,
        "Exception ID": 14,
        "Risk ID": 14,
        "Compensating Controls": 30,
        "Responsible Person": 20,
        "Target Date": 14,
        "Budget Required": 15,
    }


def get_extended_columns(sheet_type):
    """
    Return additional columns (R-X) specific to each sheet type.
    
    Args:
        sheet_type: String identifier (generation, storage, rotation, backup, certificate)
    
    Returns:
        dict: {column_name: width} for columns R onwards
    """
    extensions = {
        "generation": {
            "Random Number Generator": 25,
            "Generation Location": 20,
            "Key Ceremony Performed": 16,
            "Entropy Source Verified": 16,
            "Generation Audit Log": 16,
        },
        "storage": {
            "HSM/KMS Model": 25,
            "FIPS 140 Level": 16,
            "Access Control Method": 22,
            "Encryption at Rest": 16,
            "Physical Security": 20,
            "Key Backup Exists": 16,
        },
        "rotation": {
            "Rotation Schedule": 20,
            "Last Rotation Date": 16,
            "Next Rotation Due": 16,
            "Automated Rotation": 16,
            "Rotation Tested": 16,
            "Re-encryption Required": 18,
        },
        "backup": {
            "Backup Method": 22,
            "Backup Encryption": 16,
            "Backup Location": 22,
            "Recovery Tested": 16,
            "Last Recovery Test": 16,
            "RTO (Recovery Time)": 18,
            "Escrow Agreement": 16,
        },
        "certificate": {
            "Certificate Type": 22,
            "Issuing CA": 25,
            "Certificate Validity": 16,
            "Expiration Date": 16,
            "Auto-Renewal": 16,
            "Revocation Method": 20,
            "Monitoring Enabled": 16,
        },
    }
    
    return extensions.get(sheet_type, {})


def get_all_columns(sheet_type):
    """Combine base + extended columns for a given sheet type."""
    base = get_base_key_mgmt_columns()
    extended = get_extended_columns(sheet_type)
    base.update(extended)
    return base


# ============================================================================
# SECTION 3: GLOBAL DATA VALIDATION DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create base dropdown validations common to all key management sheets.
    Returns dict of DataValidation objects.
    """
    validations = {}
    
    # Key Type
    validations['key_type'] = DataValidation(
        type="list",
        formula1='"Encryption Key,Signing Key,Authentication Key,Certificate,Master Key,Data Encryption Key,Key Encryption Key,Session Key,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['key_type'])
    
    # Algorithm
    validations['algorithm'] = DataValidation(
        type="list",
        formula1='"AES-256,AES-128,RSA-4096,RSA-3072,RSA-2048,Ed25519,ECDSA P-256,ECDSA P-384,ChaCha20,Other,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['algorithm'])
    
    # Key Length
    validations['key_length'] = DataValidation(
        type="list",
        formula1='"256,384,512,2048,3072,4096,8192,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['key_length'])
    
    # Storage Method
    validations['storage'] = DataValidation(
        type="list",
        formula1='"HSM,Cloud KMS,TPM,Software Keystore,File-based,Database,Smart Card,Vault,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['storage'])
    
    # Data Classification
    validations['data_class'] = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['data_class'])
    
    # Lifecycle Stage
    validations['lifecycle'] = DataValidation(
        type="list",
        formula1='"Active,Suspended,Archived,Destroyed,Pending Rotation,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['lifecycle'])
    
    # Status
    validations['status'] = DataValidation(
        type="list",
        formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['status'])
    
    # Remediation Needed
    validations['remediation'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['remediation'])
    
    # Budget Required
    validations['budget'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No,⚠️ Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(validations['budget'])
    
    # Response (Yes/No/Not Applicable)
    validations['response'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No,⚠️ Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(validations['response'])
    
    # Checklist Yes/No/N/A
    validations['checklist'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['checklist'])
    
    # Exception Yes/No
    validations['exception_yn'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['exception_yn'])
    
    return validations


def create_extended_validations(ws, sheet_type):
    """Create sheet-specific dropdown validations for extended columns."""
    validations = {}
    
    if sheet_type == "generation":
        validations['rng'] = DataValidation(
            type="list",
            formula1='"Hardware RNG,OS Crypto RNG,HSM RNG,TPM RNG,FIPS 140-2 Validated,NIST SP 800-90A,Non-compliant,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['rng'])
        
        validations['gen_location'] = DataValidation(
            type="list",
            formula1='"HSM,KMS,TPM,Local System,Cloud Service,Smart Card,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['gen_location'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "storage":
        validations['fips_level'] = DataValidation(
            type="list",
            formula1='"Level 1,Level 2,Level 3,Level 4,Not Validated,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['fips_level'])
        
        validations['access_control'] = DataValidation(
            type="list",
            formula1='"Role-based,Certificate-based,MFA Required,PIN/Password,Smart Card,Dual Control,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['access_control'])
        
        validations['physical_sec'] = DataValidation(
            type="list",
            formula1='"Locked Data Center,Secure Room,Tamper-evident,Cloud Provider,Inadequate,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['physical_sec'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "rotation":
        validations['rotation_sched'] = DataValidation(
            type="list",
            formula1='"Hourly,Daily,Weekly,Monthly,Quarterly,Annually,Biennially,On Compromise,Never,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['rotation_sched'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
        
        validations['reencrypt'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,⏳ In Progress,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['reencrypt'])
    
    elif sheet_type == "backup":
        validations['backup_method'] = DataValidation(
            type="list",
            formula1='"HSM Backup,KMS Backup,Encrypted File,Key Escrow,Split Knowledge,Cloud Backup,No Backup,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['backup_method'])
        
        validations['recovery_tested'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,⏳ Scheduled,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['recovery_tested'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "certificate":
        validations['cert_type'] = DataValidation(
            type="list",
            formula1='"TLS/SSL Server,TLS/SSL Client,Code Signing,Email (S/MIME),User Authentication,CA Certificate,Wildcard,SAN,Self-signed,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_type'])
        
        validations['cert_validity'] = DataValidation(
            type="list",
            formula1='"≤90 days,91-180 days,181-365 days,366-397 days,>397 days (non-compliant),Expired,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_validity'])
        
        validations['revocation'] = DataValidation(
            type="list",
            formula1='"CRL,OCSP,OCSP Stapling,Not Configured,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['revocation'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    return validations
# ============================================================================
# SECTION 4: CHECKLIST DEFINITIONS PER KEY MANAGEMENT AREA
# ============================================================================

def get_checklist_items(sheet_type):
    """
    Return checklist items for each key management assessment type.
    
    Args:
        sheet_type: String identifier (generation, storage, rotation, backup, certificate)
    
    Returns:
        list: Checklist items as strings
    """
    checklists = {
        "generation": [
            "Approved algorithms only (AES-256, RSA ≥3072, Ed25519, ECDSA P-256+)",
            "Cryptographically secure random number generator (CSRNG) used",
            "FIPS 140-2 Level 2+ validated cryptographic module (production)",
            "Key generation performed in secure environment (HSM/KMS/TPM)",
            "Weak algorithms prohibited (DES, 3DES, MD5, SHA-1, RSA <2048)",
            "Key ceremony documented for master/root keys",
            "Dual control/split knowledge for critical key generation",
            "Key generation audit logging enabled",
            "Generated keys never exposed in plaintext outside secure boundary",
            "Entropy source quality verified and tested",
            "Key generation procedures documented",
            "Key metadata tracked (creation date, purpose, owner, classification)",
            "Test/development keys distinct from production keys",
            "Default/vendor keys replaced before production use",
            "Cryptographic policy document exists and is approved — covers algorithm selection, key lengths, use cases, and roles",
            "Approved algorithm list maintained and reviewed annually — deprecated algorithms (MD5, SHA-1, DES, 3DES, RC4) explicitly prohibited",
            "Cryptographic roles and responsibilities defined — key custodians, security officers, and approvers identified",
        ],
        "storage": [
            "Production keys stored in HSM, cloud KMS, or TPM",
            "FIPS 140-2 Level 2+ (or FIPS 140-3 equivalent) for production",
            "Keys encrypted when stored (key encryption keys protected by HSM)",
            "Keys never stored in plaintext files or code repositories",
            "Keys never stored in application configuration files",
            "Keys never stored in environment variables (production)",
            "Keys never stored in databases without encryption",
            "Access to keys restricted by role-based access control (RBAC)",
            "Multi-factor authentication required for key access",
            "Dual control/split knowledge for master keys",
            "Key access logged and monitored",
            "Keys physically secured (locked facilities, tamper-evident HSMs)",
            "Key storage location geographically documented",
            "Keys stored separately from encrypted data",
            "Backup keys stored with equal or greater security",
            "Key export restricted and logged",
            "Unused/obsolete keys securely destroyed",
        ],
        "rotation": [
            "Key rotation policy documented and approved",
            "Encryption keys (data at rest) rotated annually minimum",
            "Signing keys rotated biennially minimum",
            "Session keys rotated hourly/daily",
            "Master keys rotated every 1-3 years",
            "Key rotation automated where possible",
            "Key rotation triggers defined (time, usage, compromise)",
            "Data re-encryption performed after key rotation",
            "Old keys retained for decryption (with restricted access)",
            "Key version tracking implemented",
            "Rotation procedures tested in non-production",
            "Rotation audit logs maintained",
            "Emergency rotation procedure documented",
            "Rotation scheduling prevents service disruption",
            "Key rotation monitored and alerts configured",
            "Cryptoperiod limits enforced (max key usage duration)",
        ],
        "backup": [
            "Key backup and recovery procedures documented",
            "All critical keys have backup copies",
            "Backup keys encrypted with equal or stronger encryption",
            "Backup keys stored separately from production keys",
            "Geographic separation of backup keys (offsite/different region)",
            "Backup key access restricted (dual control/escrow)",
            "Key recovery procedures tested quarterly",
            "Recovery time objective (RTO) defined and documented",
            "Recovery point objective (RPO) defined and documented",
            "Key escrow agreements in place (where required)",
            "Split knowledge/dual control for backup key recovery",
            "Backup key inventory maintained and audited",
            "Backup keys versioned and synchronized with production",
            "Emergency recovery procedures documented",
            "Key recovery audit logging enabled",
            "Backup keys updated when production keys rotate",
            "Cloud KMS backup to different region/account",
            "HSM backup cards stored in secure facility",
        ],
        "certificate": [
            "Certificates issued from trusted public CA (for external services)",
            "Internal CA infrastructure properly secured (for internal services)",
            "Certificate validity ≤397 days (public TLS certificates)",
            "Certificate validity ≤90 days preferred (automated renewal)",
            "Self-signed certificates NOT used in production",
            "Certificate inventory maintained and up to date",
            "Certificate expiration monitoring implemented",
            "Alerts configured for 30/60/90 days before expiration",
            "Automated certificate renewal configured",
            "Certificate renewal tested in non-production",
            "Certificate revocation procedures documented",
            "CRL or OCSP configured for revocation checking",
            "Compromised certificate revocation within 24 hours",
            "Certificate private keys stored securely (HSM/KMS preferred)",
            "Certificate private keys never shared across systems",
            "Wildcard certificates use documented and restricted",
            "Certificate signing requests (CSR) process documented",
            "Certificate audit logs maintained",
            "Code signing certificates protected with hardware tokens",
            "Email certificates (S/MIME) issued per user",
            "Cryptographic import/export compliance — applicable national regulations on cryptographic products assessed (e.g. US EAR, French ANSSI, Swiss regulations)",
            "Jurisdictional restrictions identified — use of cryptographic products compliant with laws of all countries where data is processed or stored",
        ],
    }
    
    return checklists.get(sheet_type, [])


# ============================================================================
# SECTION 5: REFERENCE TABLE DEFINITIONS PER SHEET TYPE
# ============================================================================

def get_reference_tables(sheet_type):
    """
    Return reference table definitions for each sheet type.
    
    Returns:
        list: [(table_title, headers, data_rows), ...]
    """
    tables = {
        "generation": [
            (
                "APPROVED ALGORITHMS BY USE CASE",
                ["Use Case", "Minimum Algorithm", "Key Length", "Notes"],
                [
                    ("Symmetric Encryption (Data at Rest)", "AES", "256 bits", "AES-256-GCM preferred"),
                    ("Symmetric Encryption (Data in Transit)", "AES, ChaCha20", "256 bits", "ChaCha20-Poly1305 for mobile"),
                    ("Asymmetric Encryption", "RSA, ECDH", "RSA ≥3072, ECDH P-256+", "RSA-4096 for long-term"),
                    ("Digital Signatures", "RSA, Ed25519, ECDSA", "RSA ≥3072, Ed25519, ECDSA P-256+", "Ed25519 preferred for new systems"),
                    ("Key Exchange", "ECDH, RSA", "ECDH P-256+, RSA ≥3072", "Use ephemeral keys for PFS"),
                    ("Hashing", "SHA-256, SHA-384, SHA-512", "N/A (output size)", "SHA-256 minimum"),
                    ("HMAC", "HMAC-SHA256", "≥256 bits", "For message authentication"),
                ]
            ),
        ],
        "storage": [
            (
                "APPROVED KEY STORAGE SOLUTIONS",
                ["Solution Type", "Examples", "FIPS 140-2 Level", "Use Case"],
                [
                    ("Hardware HSM", "Thales Luna, Entrust nShield, Utimaco", "Level 3/4", "High-security production"),
                    ("Cloud HSM", "AWS CloudHSM, Azure Dedicated HSM", "Level 3", "Cloud-native high-security"),
                    ("Cloud KMS", "AWS KMS, Azure Key Vault, GCP Cloud KMS", "Level 2/3", "General cloud workloads"),
                    ("TPM", "TPM 2.0 (various vendors)", "Level 2", "Endpoint/device keys"),
                    ("Software Vault", "HashiCorp Vault, CyberArk", "Varies", "Non-production acceptable"),
                ]
            ),
        ],
        "rotation": [
            (
                "ROTATION SCHEDULE BY KEY TYPE",
                ["Key Type", "Minimum Rotation Frequency", "Recommended Frequency", "Trigger Events"],
                [
                    ("Master Encryption Key", "1 year", "1 year", "Compromise, personnel change"),
                    ("Data Encryption Key", "1 year", "90 days", "High volume usage, compromise"),
                    ("Key Encryption Key", "1 year", "1 year", "Master key rotation, compromise"),
                    ("Signing Key", "2 years", "1 year", "Certificate expiry, compromise"),
                    ("Session Key", "24 hours", "1 hour", "Session end, compromise"),
                    ("API Key", "90 days", "30-90 days", "Personnel change, compromise"),
                    ("TLS/SSL Certificate Private Key", "Certificate validity", "Annual", "Certificate renewal, compromise"),
                    ("SSH Key", "1 year", "90 days", "Personnel change, compromise"),
                    ("Database TDE Key", "1 year", "1 year", "Compliance requirement, compromise"),
                ]
            ),
        ],
        "backup": [
            (
                "BACKUP STRATEGIES BY KEY TYPE",
                ["Key Type", "Recommended Backup Method", "Storage Location", "Recovery Testing Frequency"],
                [
                    ("Master Keys", "HSM backup card + escrow", "Offsite secure vault", "Annually"),
                    ("Encryption Keys", "KMS automated backup", "Different region/account", "Quarterly"),
                    ("Certificate Private Keys", "Encrypted file backup", "Secure file storage", "Semi-annually"),
                    ("SSH Keys", "Encrypted backup + key management system", "Secrets vault", "Annually"),
                    ("API Keys", "Secrets management system", "Multi-region vault", "Quarterly"),
                    ("Database TDE Keys", "Native database backup + KMS", "Encrypted backup storage", "Quarterly"),
                ]
            ),
        ],
        "certificate": [
            (
                "CERTIFICATE VALIDITY STANDARDS",
                ["Certificate Type", "Maximum Validity", "Recommended Validity", "Renewal Process"],
                [
                    ("Public TLS/SSL", "397 days (13 months)", "90 days", "ACME/Let's Encrypt automated"),
                    ("Internal TLS/SSL", "2 years", "1 year", "Internal CA automated"),
                    ("Code Signing", "3 years", "1 year", "Manual with hardware token"),
                    ("Email (S/MIME)", "2 years", "1 year", "PKI automated or manual"),
                    ("Client Authentication", "2 years", "1 year", "PKI automated"),
                    ("Root CA", "20 years", "10 years", "Manual with ceremony"),
                    ("Intermediate CA", "10 years", "5 years", "Manual with ceremony"),
                ]
            ),
            (
                "CERTIFICATE LIFECYCLE STAGES",
                ["Stage", "Actions", "Responsible Role", "Documentation Required"],
                [
                    ("Request", "Generate CSR, validate domain/identity", "System Owner", "CSR details, approval"),
                    ("Issuance", "CA validation, certificate generation", "CA Admin", "Issuance record, validation log"),
                    ("Installation", "Deploy certificate to service", "System Admin", "Deployment log, verification"),
                    ("Monitoring", "Track expiration, validate configuration", "Security Team", "Monitoring dashboard"),
                    ("Renewal", "Generate new CSR, request renewal", "System Owner", "Renewal approval, testing"),
                    ("Revocation", "Revoke compromised/obsolete certificate", "CA Admin", "Incident record, revocation log"),
                    ("Archival", "Archive expired certificates", "Security Team", "Archive log, retention period"),
                ]
            ),
        ],
    }
    
    return tables.get(sheet_type, [])
# ============================================================================
# SECTION 6: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab (1–5) for applicable key management areas.', '2. Use dropdown menus for standardised entries (Status, Key Type, Algorithm, Storage Method, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Document key lifecycle details for each cryptographic system.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics per key management area.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_assessment_sheet(ws, styles, section_title, policy_ref, question, 
                           sheet_type):
    """
    Generic assessment sheet creator for key management controls.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "1. KEY GENERATION - CRYPTOGRAPHIC KEY CREATION"
        policy_ref: policy requirement text
        question: assessment question
        sheet_type: key for columns/checklist/tables (generation, storage, rotation, backup, certificate)
    """
    columns = get_all_columns(sheet_type)
    checklist_items = get_checklist_items(sheet_type)
    reference_tables = get_reference_tables(sheet_type)
    
    total_cols = len(columns)
    last_col_letter = get_column_letter(total_cols)
    
    # ---------- HEADER ----------
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{section_title}\nPolicy Requirement: {policy_ref}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.8.24: Use of Cryptography"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ---------- ASSESSMENT QUESTION ----------
    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ---------- RESPONSE DROPDOWN ----------
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    # Create validations
    base_validations = create_base_validations(ws)
    extended_validations = create_extended_validations(ws, sheet_type)
    base_validations['response'].add(ws["B4"])

    # ---------- COLUMN HEADERS ----------
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ---------- EXAMPLE ROW ----------
    example_row = 7
    example_values = _get_example_row(sheet_type, total_cols)
    for col_idx, value in enumerate(example_values, start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------- DATA ENTRY ROWS (8-20) ----------
    start_row = 8
    end_row = 20
    
    for r in range(start_row, end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Apply base validations to standard columns (B-Q, indices 2-17)
    for r in range(start_row, end_row + 1):
        base_validations['key_type'].add(ws.cell(row=r, column=2))      # B: Key Type
        base_validations['algorithm'].add(ws.cell(row=r, column=3))     # C: Algorithm
        base_validations['key_length'].add(ws.cell(row=r, column=4))    # D: Key Length
        base_validations['storage'].add(ws.cell(row=r, column=5))       # E: Storage Method
        base_validations['data_class'].add(ws.cell(row=r, column=6))    # F: Data Classification
        base_validations['lifecycle'].add(ws.cell(row=r, column=7))     # G: Lifecycle Stage
        base_validations['status'].add(ws.cell(row=r, column=8))        # H: Status
        base_validations['remediation'].add(ws.cell(row=r, column=11))  # K: Remediation Needed
        base_validations['budget'].add(ws.cell(row=r, column=17))       # Q: Budget Required

    # Apply extended validations for columns R onwards (18+)
    _apply_extended_validations(ws, sheet_type, extended_validations, start_row, end_row)

    ws.freeze_panes = "A7"

    next_row = end_row + 2

    # ---------- COMPLIANCE CHECKLIST ----------
    ws[f"A{next_row}"] = f"{section_title.split('-')[0].strip().upper()} CHECKLIST"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)
    next_row += 1

    ws[f"A{next_row}"] = "☐"
    ws[f"B{next_row}"] = "Requirement"
    ws[f"C{next_row}"] = "Status"
    for col in ["A", "B", "C"]:
        ws[f"{col}{next_row}"].font = Font(bold=True)
    next_row += 1

    checklist_start = next_row
    for item in checklist_items:
        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = item
        ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{next_row}"].border = styles["border"]

        base_validations['checklist'].add(ws[f"C{next_row}"])
        next_row += 1

    # Checklist score
    ws[f"A{next_row}"] = "Checklist Score:"
    ws[f"A{next_row}"].font = Font(bold=True)
    ws[f"B{next_row}"] = (
        f'=IFERROR(ROUND(COUNTIF(C{checklist_start}:C{next_row-1},"✅ Yes")/'
        f'COUNTA(C{checklist_start}:C{next_row-1})*100,0)&"%","0%")'
    )
    ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
    next_row += 2

    # ---------- REFERENCE TABLES ----------
    for table_title, headers, data_rows in reference_tables:
        ws[f"A{next_row}"] = table_title
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1
        
        # Table headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=header)
            cell.font = styles["column_header"]["font"]
            cell.fill = styles["column_header"]["fill"]
            cell.alignment = styles["column_header"]["alignment"]
            cell.border = styles["column_header"]["border"]
        next_row += 1
        
        # Table data rows
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = styles["border"]
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            next_row += 1
        
        next_row += 1  # Space between tables

    # ---------- EXCEPTION/DEVIATION BLOCK ----------
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "EXCEPTION / DEVIATION DOCUMENTATION (Complete if Status = Partial or Non-Compliant)"
    ws[f"A{next_row}"].font = styles["exception_header"]["font"]
    ws[f"A{next_row}"].fill = styles["exception_header"]["fill"]
    ws[f"A{next_row}"].alignment = styles["exception_header"]["alignment"]

    next_row += 1
    exception_fields = [
        ("Formal exception request submitted:", "✅ Yes,❌ No"),
        ("Exception ID:", ""),
        ("Risk acceptance documented:", "✅ Yes,❌ No"),
        ("Risk ID:", ""),
        ("Compensating Controls (summary):", ""),
        ("☐ Enhanced monitoring and alerting", ""),
        ("☐ Restricted network access / air-gapped systems", ""),
        ("☐ Additional encryption layer", ""),
        ("☐ Manual key management procedures", ""),
        ("☐ Increased audit frequency", ""),
        ("☐ Physical security controls", ""),
        ("☐ Other (describe):", ""),
    ]

    for label, options in exception_fields:
        ws[f"A{next_row}"] = label
        ws[f"A{next_row}"].font = Font(bold=True)
        ws.merge_cells(f"B{next_row}:D{next_row}")
        ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{next_row}"].border = styles["border"]
        ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]

        if options:
            base_validations['exception_yn'].add(ws[f"B{next_row}"])

        next_row += 1

    # ---------- NOTES ----------
    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "ADDITIONAL NOTES / COMMENTS"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)

    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row+8}")
    ws[f"A{next_row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{next_row}"].border = styles["border"]
    ws[f"A{next_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths for checklist/tables
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 20

    # GS-AS-011: Borders on all merged cell ranges
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for _mr in list(ws.merged_cells.ranges):
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                ws.cell(row=_r, column=_c).border = _border

    return (start_row, end_row, 8)  # Return status column index (H = 8)


def _get_example_row(sheet_type, total_cols):
    """Generate example row values based on sheet type."""
    examples = {
        "generation": [
            "Example: Production Database Encryption - FINDB01", "Data Encryption Key", "AES-256", "256",
            "HSM", "Restricted", "Active", "✅ Compliant",
            "/evidence/key-mgmt/FINDB01-key-generation-log.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "HSM RNG", "HSM", "Yes", "Yes", "Yes"
        ],
        "storage": [
            "Example: Azure Key Vault - App1 Keys", "Key Encryption Key", "RSA-4096", "4096",
            "Cloud KMS", "Confidential", "Active", "✅ Compliant",
            "/evidence/key-mgmt/app1-kv-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Azure Key Vault", "Level 2", "MFA Required", "Yes", "Cloud Provider", "Yes"
        ],
        "rotation": [
            "Example: API Signing Keys", "Signing Key", "RSA-3072", "3072",
            "Cloud KMS", "Confidential", "Active", "✅ Compliant",
            "/evidence/key-mgmt/api-key-rotation.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Quarterly", "2024-10-01", "2025-01-01", "Yes", "Yes", "No"
        ],
        "backup": [
            "Example: Master Encryption Key Backup", "Master Key", "AES-256", "256",
            "HSM", "Restricted", "Active", "✅ Compliant",
            "/evidence/key-mgmt/master-key-backup.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "HSM Backup", "Yes", "Offsite Secure Vault", "Yes", "2024-06-15", "<1 hour", "Yes"
        ],
        "certificate": [
            "Example: Production Web Server TLS", "Certificate", "RSA-2048", "2048",
            "TPM", "Public", "Active", "✅ Compliant",
            "/evidence/key-mgmt/web-tls-cert.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "TLS/SSL Server", "Let's Encrypt", "≤90 days", "2025-03-15", "Yes", "OCSP", "Yes"
        ],
    }
    
    base_example = examples.get(sheet_type, ["Example"] * total_cols)
    # Pad or trim to match total_cols
    if len(base_example) < total_cols:
        base_example.extend(["N/A"] * (total_cols - len(base_example)))
    return base_example[:total_cols]


def _apply_extended_validations(ws, sheet_type, validations, start_row, end_row):
    """Apply extended column validations based on sheet type."""
    if sheet_type == "generation":
        for r in range(start_row, end_row + 1):
            validations['rng'].add(ws.cell(row=r, column=18))        # R
            validations['gen_location'].add(ws.cell(row=r, column=19)) # S
            validations['yn_na'].add(ws.cell(row=r, column=20))      # T: Key Ceremony
            validations['yn_na'].add(ws.cell(row=r, column=21))      # U: Entropy Verified
            validations['yn_na'].add(ws.cell(row=r, column=22))      # V: Audit Log
    
    elif sheet_type == "storage":
        for r in range(start_row, end_row + 1):
            # R: HSM/KMS Model - free text
            validations['fips_level'].add(ws.cell(row=r, column=19))    # S
            validations['access_control'].add(ws.cell(row=r, column=20)) # T
            validations['yn_na'].add(ws.cell(row=r, column=21))         # U: Encryption at Rest
            validations['physical_sec'].add(ws.cell(row=r, column=22))  # V
            validations['yn_na'].add(ws.cell(row=r, column=23))         # W: Key Backup Exists
    
    elif sheet_type == "rotation":
        for r in range(start_row, end_row + 1):
            validations['rotation_sched'].add(ws.cell(row=r, column=18)) # R
            # S: Last Rotation Date - date picker
            # T: Next Rotation Due - date picker
            validations['yn_na'].add(ws.cell(row=r, column=21))         # U: Automated Rotation
            validations['yn_na'].add(ws.cell(row=r, column=22))         # V: Rotation Tested
            validations['reencrypt'].add(ws.cell(row=r, column=23))     # W
    
    elif sheet_type == "backup":
        for r in range(start_row, end_row + 1):
            validations['backup_method'].add(ws.cell(row=r, column=18)) # R
            validations['yn_na'].add(ws.cell(row=r, column=19))         # S: Backup Encryption
            # T: Backup Location - free text
            validations['recovery_tested'].add(ws.cell(row=r, column=21)) # U
            # V: Last Recovery Test - date picker
            # W: RTO - free text
            validations['yn_na'].add(ws.cell(row=r, column=24))         # X: Escrow Agreement
    
    elif sheet_type == "certificate":
        for r in range(start_row, end_row + 1):
            validations['cert_type'].add(ws.cell(row=r, column=18))     # R
            # S: Issuing CA - free text
            validations['cert_validity'].add(ws.cell(row=r, column=20)) # T
            # U: Expiration Date - date picker
            validations['yn_na'].add(ws.cell(row=r, column=22))         # V: Auto-Renewal
            validations['revocation'].add(ws.cell(row=r, column=23))    # W
            validations['yn_na'].add(ws.cell(row=r, column=24))         # X: Monitoring Enabled
# ============================================================================
# SECTION 8: INDIVIDUAL ASSESSMENT SHEET DEFINITIONS
# ============================================================================

def create_1_key_generation(ws, styles):
    """1. Key Generation - matches markdown spec exactly."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1. KEY GENERATION - CRYPTOGRAPHIC KEY CREATION",
        policy_ref="Cryptographic keys MUST be generated using approved algorithms with cryptographically secure random number generators (Policy Section 2.4.1)",
        question="Does your organisation generate cryptographic keys for encryption, signing, or authentication purposes?",
        sheet_type="generation",
    )


def create_2_key_storage(ws, styles):
    """2. Key Storage - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2. KEY STORAGE - SECURE KEY PROTECTION",
        policy_ref="Cryptographic keys MUST be stored in hardware security modules (HSM), cloud KMS, or TPM for production systems. Software keystores permitted only for non-production with documented risk acceptance (Policy Section 2.4.2)",
        question="Does your organisation store cryptographic keys that require protection?",
        sheet_type="storage",
    )


def create_3_key_rotation(ws, styles):
    """3. Key Rotation - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="3. KEY ROTATION - CRYPTOGRAPHIC KEY LIFECYCLE MANAGEMENT",
        policy_ref="Cryptographic keys MUST be rotated according to defined schedules: Encryption keys annually, signing keys biennially, session keys hourly/daily (Policy Section 2.4.3)",
        question="Does your organisation have a key rotation policy and implementation for cryptographic keys?",
        sheet_type="rotation",
    )


def create_4_key_backup_recovery(ws, styles):
    """4. Key Backup & Recovery - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4. KEY BACKUP & RECOVERY - BUSINESS CONTINUITY",
        policy_ref="Cryptographic keys MUST have documented backup and recovery procedures. Backup keys stored with equal or greater security than production keys (Policy Section 2.4.4)",
        question="Does your organisation maintain backup and recovery procedures for cryptographic keys?",
        sheet_type="backup",
    )


def create_5_certificate_management(ws, styles):
    """5. Certificate Management - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="5. CERTIFICATE MANAGEMENT - DIGITAL CERTIFICATE LIFECYCLE",
        policy_ref="Digital certificates MUST be issued from trusted Certificate Authorities. Certificate lifecycle management includes issuance, renewal, revocation, and expiration monitoring (Policy Section 2.4.5)",
        question="Does your organisation use digital certificates for TLS/SSL, code signing, email signing, or authentication?",
        sheet_type="certificate",
    )


def create_6_key_destruction(ws, styles):
    """6. Key Destruction - ISO A.8.24 requires documented key destruction procedures."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_names = [
        "Key ID", "Key Type", "Algorithm", "Creation Date",
        "Scheduled Destruction Date", "Destruction Method", "Actual Destruction Date",
        "Verification Method", "Authorised By", "Status",
    ]
    col_widths = [16, 20, 18, 16, 24, 28, 24, 28, 22, 18]

    # --- Header row 1 ---
    last_col = get_column_letter(len(col_names))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = ("6. KEY DESTRUCTION - CRYPTOGRAPHIC KEY DISPOSAL\n"
                "Policy Requirement: Cryptographic keys that have reached end-of-life "
                "MUST be securely destroyed and destruction documented (Policy Section 2.4.6)")
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # --- Row 2: ISO reference ---
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.8.24: Use of Cryptography"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Row 3: Assessment question ---
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = "Does your organisation have documented procedures for secure cryptographic key destruction and disposal?"
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # --- Row 4: Response dropdown ---
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = border

    dv_resp = DataValidation(
        type="list",
        formula1='"Yes,No,Partially,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_resp)
    dv_resp.add(ws["B4"])

    # --- Row 6: Column headers ---
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(zip(col_names, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # --- Row 7: Sample row (F2F2F2 grey) ---
    sample_row = 7
    sample_vals = [
        "KEY-001", "Data Encryption Key", "AES-256", "01.01.2023",
        "01.01.2025", "Crypto-erasure", "15.12.2024",
        "Third-party audit + log", "CISO", "✅ Destroyed",
    ]
    for col_idx, val in enumerate(sample_vals, start=1):
        cell = ws.cell(row=sample_row, column=col_idx, value=val)
        cell.font = Font(italic=True, color="808080")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    # --- Data Validations ---
    dv_key_type = DataValidation(
        type="list",
        formula1='"Encryption Key,Signing Key,Authentication Key,Certificate,Master Key,Data Encryption Key,Key Encryption Key,Session Key,N/A"',
        allow_blank=True
    )
    dv_algorithm = DataValidation(
        type="list",
        formula1='"AES-256,AES-128,RSA-4096,RSA-3072,RSA-2048,Ed25519,ECDSA P-256,ECDSA P-384,ChaCha20,Other,N/A"',
        allow_blank=True
    )
    dv_destr_method = DataValidation(
        type="list",
        formula1='"Hardware Shredding,Crypto-erasure,Physical Destruction,Vendor-managed,Degaussing,Overwrite (multi-pass),N/A"',
        allow_blank=True
    )
    dv_status = DataValidation(
        type="list",
        formula1='"\u2705 Destroyed,\u23f3 Pending,\u274c Overdue,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_key_type)
    ws.add_data_validation(dv_algorithm)
    ws.add_data_validation(dv_destr_method)
    ws.add_data_validation(dv_status)

    # --- Rows 8-57: Data entry rows (FFFFCC) — 50 rows standard per DS-005 ---
    start_row = 8
    end_row = 57
    for r in range(start_row, end_row + 1):
        for c in range(1, len(col_names) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border
            cell.alignment = styles["input_cell"]["alignment"]
        dv_key_type.add(ws.cell(row=r, column=2))    # B: Key Type
        dv_algorithm.add(ws.cell(row=r, column=3))   # C: Algorithm
        dv_destr_method.add(ws.cell(row=r, column=6))  # F: Destruction Method
        dv_status.add(ws.cell(row=r, column=10))     # J: Status

    ws.freeze_panes = "A7"

    # --- Compliance Checklist ---
    next_row = end_row + 2
    ws[f"A{next_row}"] = "KEY DESTRUCTION CHECKLIST"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)
    next_row += 1

    ws[f"A{next_row}"] = "☐"
    ws[f"B{next_row}"] = "Requirement"
    ws[f"C{next_row}"] = "Status"
    for col in ["A", "B", "C"]:
        ws[f"{col}{next_row}"].font = Font(bold=True)
    next_row += 1

    checklist_items = [
        "Key destruction policy documented and approved",
        "Destruction procedures defined for each key type (HSM, software, cloud)",
        "Cryptographic erasure used where physical destruction not possible",
        "Destruction authority (authorisation) defined and enforced",
        "Dual-person authorisation required for master/critical key destruction",
        "Destruction records maintained (what, when, who, how)",
        "Third-party verification for high-sensitivity key destruction",
        "Hardware-based keys: HSM/token physically destroyed or cryptographically zeroed",
        "Cloud KMS keys: provider deletion confirmation obtained",
        "Destruction audit trail retained for minimum 7 years",
        "All copies of key (backup, escrow) destroyed simultaneously",
        "Destruction tested and verified in non-production",
        "Key destruction triggers defined (expiry, compromise, personnel change)",
        "Emergency destruction procedure documented",
        "Post-destruction verification procedure in place",
    ]

    dv_checklist = DataValidation(
        type="list",
        formula1='"\u2705 Yes,\u274c No,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_checklist)

    checklist_start = next_row
    for item in checklist_items:
        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = item
        ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{next_row}"].border = border
        dv_checklist.add(ws[f"C{next_row}"])
        next_row += 1

    ws[f"A{next_row}"] = "Checklist Score:"
    ws[f"A{next_row}"].font = Font(bold=True)
    ws[f"B{next_row}"] = (
        f'=IFERROR(ROUND(COUNTIF(C{checklist_start}:C{next_row-1},"\u2705 Yes")/'
        f'COUNTA(C{checklist_start}:C{next_row-1})*100,0)&"%","0%")'
    )
    ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
    next_row += 2

    # GS-AS-011: Borders on all merged cell ranges
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for _mr in list(ws.merged_cells.ranges):
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                ws.cell(row=_r, column=_c).border = _border


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD (11 ANALYSIS SECTIONS)
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard — TABLE 1/2/3, all tables uniform 7-col A:G width."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Row 1: Title ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font      = styles["header"]["font"]
    ws["A1"].fill      = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border

    # --- Row 2: ISO reference ---
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.8.24 | Cryptographic Key Management"
    ws["A2"].font      = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    # --- TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW ---
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    c.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="003366")
    #c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    for col_idx, h in enumerate(["Assessment Area", "Total Items", "Compliant", "Partial",
                                  "Non-Compliant", "N/A", "Compliance %"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill      = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    areas = [
        ("1. Key Generation",         "1. Key Generation"),
        ("2. Key Storage",            "2. Key Storage"),
        ("3. Key Rotation",           "3. Key Rotation"),
        ("4. Key Backup & Recovery",  "4. Key Backup & Recovery"),
        ("5. Certificate Management", "5. Certificate Management"),
        ("6. Key Destruction",        "6. Key Destruction"),
    ]
    row += 1
    start_data_row = row
    for label, sheet in areas:
        ws.cell(row=row, column=1, value=label)
        rng = f"'{sheet}'!H8:H90"
        ws.cell(row=row, column=2, value=f"=COUNTA({rng})")
        ws.cell(row=row, column=3, value=f'=COUNTIF({rng},"✅*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({rng},"⚠️*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({rng},"❌*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({rng},"N/A")')
        ws.cell(row=row, column=7,
                value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font  = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill                 = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border               = border
    for col in range(2, 7):
        c       = ws.cell(row=total_row, column=col)
        c.value = f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{total_row - 1})"
        c.font  = Font(name="Calibri", size=10, bold=True)
        c.fill  = PatternFill("solid", fgColor="D9D9D9")
        c.border = border
    c       = ws.cell(row=total_row, column=7)
    c.value = (f'=IF((B{total_row}-F{total_row})=0,"0%",'
               f'ROUND(C{total_row}/(B{total_row}-F{total_row})*100,1)&"%")')
    c.font   = Font(name="Calibri", size=12, bold=True, color="000000")
    c.fill   = PatternFill("solid", fgColor="D9D9D9")
    c.border = border

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 15

    _kpis = [
        ("Total Cryptographic Keys Managed (count)", "Count"),
        ("Keys in HSM / TPM / Cloud KMS", "≥90%"),
        ("Approved Algorithm Compliance – AES-256 / RSA-3072+", "100%"),
        ("Automated Key Rotation Coverage", "≥80%"),
        ("Certificates Expiring < 30 Days (count)", "0"),
        ("Expired / Non-Compliant Certificates (count)", "0"),
        ("FIPS 140-2 Level 2+ Validated Modules", "100%"),
        ("Key Backup & Recovery Tested", "100%"),
        ("Certificate Auto-Renewal Enabled", "≥90%"),
        ("Evidence Documentation Rate", "100%"),
    ]
    # --- TABLE 2: KEY METRICS ---
    row = total_row + 2
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    #c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    for col_idx, h in enumerate(["KPI", "Current Value", "Target", "Status",
                                  "Last Updated", "Owner", "Notes"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    dv_kpi_sts = DataValidation(type="list",
        formula1='"\u2705 On Target,\u26a0\ufe0f At Risk,\u274c Below Target"',
        allow_blank=True)
    ws.add_data_validation(dv_kpi_sts)

    # --- TABLE 2a: AUTO-COMPUTED METRICS (formula-driven) ---
    row += 1
    t2a_headers = ["Metric (Auto-computed)", "Value", "Notes", "", "", "", ""]
    for col_idx, h in enumerate(t2a_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h if h else "")
        if h:
            c.font = Font(name="Calibri", size=10, bold=True, color="000000")
            c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border

    row += 1
    # Key mgmt sheets: C=Algorithm, D=Key Length, G=Lifecycle Stage, H=Status
    _km_sheets = [
        "1. Key Generation", "2. Key Storage", "3. Key Rotation",
        "4. Key Backup & Recovery", "5. Certificate Management", "6. Key Destruction",
    ]
    def _sum_countif_km(col_letter, value):
        parts = [f'COUNTIF(\'{s}\'!{col_letter}8:{col_letter}90,\"{value}\")' for s in _km_sheets]
        return "=SUM(" + ",".join(parts) + ")"

    auto_metrics = [
        ("Non-Compliant Items (all domains)",
         "=" + "+".join([f'COUNTIF(\'{s}\'!H8:H90,"❌*")' for s in _km_sheets])),
        ("Partial Items (all domains)",
         "=" + "+".join([f'COUNTIF(\'{s}\'!H8:H90,"⚠️*")' for s in _km_sheets])),
        ("Compliant Items (all domains)",
         "=" + "+".join([f'COUNTIF(\'{s}\'!H8:H90,"✅*")' for s in _km_sheets])),
        ("Total Assessed Items (all domains)",
         "=" + "+".join([f'COUNTA(\'{s}\'!H8:H90)' for s in _km_sheets])),
        ("AES-256 Keys (count)",           _sum_countif_km("C", "AES-256")),
        ("RSA-4096 Keys (count)",          _sum_countif_km("C", "RSA-4096")),
        ("RSA-3072 Keys (count)",          _sum_countif_km("C", "RSA-3072")),
        ("Non-Approved Algorithms (count)", _sum_countif_km("C", "Other")),
        ("Active Keys (Lifecycle: Active)", _sum_countif_km("G", "Active")),
        ("Archived Keys (Lifecycle: Archived)", _sum_countif_km("G", "Archived")),
        ("Destroyed Keys (Lifecycle: Destroyed)", _sum_countif_km("G", "Destroyed")),
        ("Keys Pending Rotation",          _sum_countif_km("G", "Pending Rotation")),
    ]
    for metric, fml in auto_metrics:
        ws.cell(row=row, column=1, value=metric).border = border
        cx = ws.cell(row=row, column=2, value=fml)
        cx.font   = Font(name="Calibri", size=10, color="000000")
        cx.border = border
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    row += 1  # spacer before manual KPIs

    row += 1
    for kpi, target in _kpis:
        ws.cell(row=row, column=1, value=kpi).border = border
        c2 = ws.cell(row=row, column=2)
        c2.border = border
        ws.cell(row=row, column=3, value=target).border = border
        c4 = ws.cell(row=row, column=4)
        c4.border = border
        dv_kpi_sts.add(c4)
        for col in range(5, 8):
            ws.cell(row=row, column=col).border = border
        row += 1
    r = row

    # 2 bordered buffer rows + blank gap before TABLE 3 (Gold Standard)
    for _buf in range(2):
        for _col in range(1, 8):
            ws.cell(row=r, column=_col).border = border
        r += 1

    # --- TABLE 3: DOMAIN COMPLIANCE SUMMARY ---
    row = r + 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: DOMAIN COMPLIANCE SUMMARY")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    t3_headers = [
        "Key Management Domain",
        "Non-Compliant",
        "Partial",
        "Compliant",
        "N/A",
        "Total Items",
        "Compliance %",
    ]
    for col_idx, h in enumerate(t3_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill      = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    # Key mgmt domains: status col H (col 8), data rows 8-90
    t3_domains = [
        ("1. Key Generation",         "1. Key Generation"),
        ("2. Key Storage",            "2. Key Storage"),
        ("3. Key Rotation",           "3. Key Rotation"),
        ("4. Key Backup & Recovery",  "4. Key Backup & Recovery"),
        ("5. Certificate Management", "5. Certificate Management"),
        ("6. Key Destruction",        "6. Key Destruction"),
    ]
    row += 1
    t3_start = row
    for label, sheet in t3_domains:
        rng    = f'\'{sheet}\'!H8:H90'
        non_c  = f'=COUNTIF({rng},"❌*")'
        part   = f'=COUNTIF({rng},"⚠️*")'
        compl  = f'=COUNTIF({rng},"✅*")'
        na_    = f'=COUNTIF({rng},"N/A")'
        total_ = f'=COUNTA({rng})'
        pct_r  = row
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font   = Font(name="Calibri", size=10, color="000000")
        c1.fill   = PatternFill("solid", fgColor="FFFFCC")
        c1.border = border
        for col_idx, fml in enumerate([non_c, part, compl, na_, total_], 2):
            cx = ws.cell(row=row, column=col_idx, value=fml)
            cx.font   = Font(name="Calibri", size=10, color="000000")
            cx.fill   = PatternFill("solid", fgColor="FFFFCC")
            cx.border = border
        c7 = ws.cell(row=row, column=7,
                     value=f'=IF((F{pct_r}-E{pct_r})=0,0,D{pct_r}/(F{pct_r}-E{pct_r}))')
        c7.font          = Font(name="Calibri", size=10, color="000000")
        c7.fill          = PatternFill("solid", fgColor="FFFFCC")
        c7.number_format = "0.0%"
        c7.border        = border
        row += 1

    # TOTAL row
    t3_end = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font  = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=1).fill   = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=row, column=1).border = border
    for col in range(2, 7):
        cx = ws.cell(row=row, column=col)
        cx.value  = f"=SUM({get_column_letter(col)}{t3_start}:{get_column_letter(col)}{t3_end})"
        cx.font   = Font(name="Calibri", size=10, bold=True)
        cx.fill   = PatternFill("solid", fgColor="D9D9D9")
        cx.border = border
    tot_row = row
    cx = ws.cell(row=row, column=7,
                 value=f'=IF((F{tot_row}-E{tot_row})=0,0,D{tot_row}/(F{tot_row}-E{tot_row}))')
    cx.font          = Font(name="Calibri", size=10, bold=True, color="000000")
    cx.fill          = PatternFill("solid", fgColor="D9D9D9")
    cx.number_format = "0.0%"
    cx.border        = border
    row += 1

    # Borders on all merged cell ranges (GS-AS-011)
    _t = Side(style="thin")
    _b = Border(left=_t, right=_t, top=_t, bottom=_t)
    for _mr in list(ws.merged_cells.ranges):
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                ws.cell(row=_r, column=_c).border = _b

    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create evidence register for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validation dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Key generation log,HSM configuration,KMS architecture,Rotation log,Backup procedure,Recovery test,Certificate inventory,CA audit report,FIPS certificate,Key ceremony record,Access log,Destruction record,Algorithm verification,Third-party audit,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    # Sample row (row 5) — F2F2F2 grey with realistic example data
    _er_sample = [
        "EV-001", "1. Key Generation", "Configuration file",
        "TLS/encryption configuration export for audit evidence",
        "/evidence/a824/config-export-2025.txt",
        "01.01.2025", "Security Team", "✅ Verified",
    ]
    for c, val in enumerate(_er_sample, start=1):
        cell = ws.cell(row=5, column=c, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if c == 1:
            cell.font = Font(italic=True, color="555555")
    dv_type.add(ws.cell(row=5, column=3))
    dv_ver.add(ws.cell(row=5, column=8))

    # Empty data rows (rows 6-105) — 100 FFFFCC input rows, NO pre-filled IDs
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(ws):
    """Create approval and sign-off sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.24.4 - Key Management Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "='Summary Dashboard'!G11"),
        ("Assessment Status:", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    for merge_range in list(ws.merged_cells.ranges):
        for _row in range(merge_range.min_row, merge_range.max_row + 1):
            for _col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=_row, column=_col).border = border
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def main() -> int:
    """Main execution function - orchestrates workbook creation."""
    try:
        logger.info("=" * 78)
        logger.info("ISMS-IMP-A.8.24.4 - Key Management Assessment Generator")
        logger.info("ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography (Key Management)")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/9] Creating Instructions & Legend sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/9] Creating 1. Key Generation sheet...")
        create_1_key_generation(wb["1. Key Generation"], styles)

        logger.info("[3/9] Creating 2. Key Storage sheet...")
        create_2_key_storage(wb["2. Key Storage"], styles)

        logger.info("[4/9] Creating 3. Key Rotation sheet...")
        create_3_key_rotation(wb["3. Key Rotation"], styles)

        logger.info("[5/9] Creating 4. Key Backup & Recovery sheet...")
        create_4_key_backup_recovery(wb["4. Key Backup & Recovery"], styles)

        logger.info("[6/9] Creating 5. Certificate Management sheet...")
        create_5_certificate_management(wb["5. Certificate Management"], styles)

        logger.info("[7/11] Creating 6. Key Destruction sheet...")
        create_6_key_destruction(wb["6. Key Destruction"], styles)

        logger.info("[8/11] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

        logger.info("[9/11] Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[10/11] Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        output_path = _wkbk_dir / OUTPUT_FILENAME
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"SUCCESS: {output_path}")
        logger.info("Workbook Structure:")
        logger.info("  - Instructions & Legend - Document info and guidance")
        logger.info("  - 5 Assessment Sheets:")
        logger.info("    1. Key Generation (14 checklist items + 1 reference table)")
        logger.info("    2. Key Storage (17 checklist items + 1 reference table)")
        logger.info("    3. Key Rotation (16 checklist items + 1 reference table)")
        logger.info("    4. Key Backup & Recovery (18 checklist items + 1 reference table)")
        logger.info("    5. Certificate Management (20 checklist items + 2 reference tables)")
        logger.info("  - Summary Dashboard - 11 comprehensive analysis sections:")
        logger.info("    - Compliance summary by key management area")
        logger.info("    - Key management maturity assessment")
        logger.info("    - Algorithm distribution analysis")
        logger.info("    - Storage method distribution with security ratings")
        logger.info("    - Key rotation coverage analysis")
        logger.info("    - Certificate management health metrics")
        logger.info("    - FIPS 140-2 validation status")
        logger.info("    - Critical gaps requiring immediate attention (8 rows)")
        logger.info("    - Risk summary by data classification")
        logger.info("    - Key backup coverage metrics")
        logger.info("    - Overall assessment summary with maturity level")
        logger.info("  - Evidence Register - 100 evidence entry rows")
        logger.info("  - Approval Sign-Off - Complete workflow with 3 sign-off levels")

        logger.info("Column Structure:")
        logger.info("  - Base columns (A-Q): 17 standard columns across all sheets")
        logger.info("  - Extended columns (R-X):")
        logger.info("    - Key Generation: +5 columns (RNG, location, ceremony, entropy, audit)")
        logger.info("    - Key Storage: +6 columns (HSM/KMS, FIPS level, access control, etc.)")
        logger.info("    - Key Rotation: +6 columns (schedule, dates, automation, testing, etc.)")
        logger.info("    - Key Backup & Recovery: +7 columns (method, encryption, location, RTO, etc.)")
        logger.info("    - Certificate Management: +7 columns (type, CA, validity, renewal, etc.)")

        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions & Legend")
        logger.info("  2) Fill yellow cells in each assessment sheet (1-5)")
        logger.info("  3) Check compliance checklists per key management area:")
        logger.info("     - Key Generation: Approved algorithms, CSRNG, FIPS validation")
        logger.info("     - Key Storage: HSM/KMS usage, FIPS 140-2 Level 2+, access controls")
        logger.info("     - Key Rotation: Rotation schedules, automation, testing")
        logger.info("     - Key Backup & Recovery: Backup procedures, recovery testing, escrow")
        logger.info("     - Certificate Management: CA trust, validity periods, revocation")
        logger.info("  4) Review reference tables (approved algorithms, storage solutions, etc.)")
        logger.info("  5) Document exceptions/deviations as needed")
        logger.info("  6) Maintain Evidence Register entries")
        logger.info("  7) Review Summary Dashboard:")
        logger.info("     - Overall compliance by key management domain")
        logger.info("     - Maturity assessment per area")
        logger.info("     - Algorithm and storage method distribution")
        logger.info("     - Key rotation and backup coverage")
        logger.info("     - Certificate health and FIPS validation status")
        logger.info("     - Critical gaps and risk summary")
        logger.info("  8) Complete multi-level approval sign-off workflow")
        logger.info("=" * 78)
        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
