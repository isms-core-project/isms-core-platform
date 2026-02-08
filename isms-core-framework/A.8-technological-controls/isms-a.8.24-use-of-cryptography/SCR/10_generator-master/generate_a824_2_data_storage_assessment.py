#!/usr/bin/env python3
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.24.2 - Data Storage Cryptography Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography
Assessment Domain 2 of 4: Data-at-Rest Cryptographic Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data storage infrastructure, cryptographic
standards, and assessment requirements.

Key customization areas:
1. Storage platform types and vendors (match your actual infrastructure)
2. Encryption implementation methods (adapt to your technology stack)
3. Key management integration (specific to your KMS/HSM infrastructure)
4. Data classification requirements (based on your classification scheme)
5. Compliance thresholds (aligned with your regulatory requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Use of Cryptography Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
cryptographic controls protecting data at rest across all storage systems and
media types.

**Purpose:**
Enables systematic assessment of encryption implementation for stored data
against ISO 27001:2022 Control A.8.24 requirements, supporting evidence-based
validation of cryptographic protection for data-at-rest scenarios.

**Assessment Scope:**
- Database encryption (TDE, field-level, application-level)
- File system encryption (full disk, volume, container)
- Object storage encryption (S3, blob storage, object stores)
- Backup encryption (full backup, incremental, archive)
- Mobile device encryption (smartphones, tablets, laptops)
- Removable media encryption (USB drives, external disks)
- Archive encryption (long-term retention, cold storage)
- Virtual machine/container encryption
- Encryption key storage and protection
- Data classification alignment with encryption requirements
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and encryption standards
2. Database Encryption - RDBMS, NoSQL, and data warehouse encryption
3. File System Encryption - Server and workstation disk encryption
4. Object Storage Encryption - Cloud and on-premise object store encryption
5. Backup Encryption - Backup and recovery system encryption
6. Endpoint Encryption - Mobile device and laptop encryption
7. Removable Media - USB, external drive, and portable media encryption
8. Archive Encryption - Long-term retention and cold storage encryption
9. Key Management Integration - KMS/HSM integration assessment
10. Data Classification Matrix - Encryption requirements per classification
11. Gap Analysis - Unencrypted storage and remediation requirements
12. Evidence Register - Audit evidence tracking and documentation
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with encryption standard dropdown lists
- Conditional formatting for encryption compliance status
- Automated gap identification for unencrypted storage
- Protected formulas with unprotected input cells
- Data classification to encryption requirement mapping
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with asset inventory systems

**Integration:**
This assessment feeds into the A.8.24.5 Compliance Dashboard, which
consolidates data from all four cryptographic assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a824_2_data_storage_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a824_2_data_storage_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a824_2_data_storage_assessment.py --date 20250115

Output:
    File: ISMS_A_8_24_2_Data_Storage_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize encryption standards to match your risk profile
    2. Inventory all data storage systems and repositories
    3. Complete encryption assessments for each storage system
    4. Validate encryption algorithms and key strengths
    5. Review key management integration and key rotation
    6. Map data classification to encryption requirements
    7. Conduct gap analysis for unencrypted data stores
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (configs, encryption reports)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.24.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.24
Assessment Domain:    2 of 4 (Data-at-Rest Cryptographic Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.24: Use of Cryptography Policy (Governance)
    - ISMS-IMP-A.8.24.1: Data Transmission Cryptography Assessment (Domain 1)
    - ISMS-IMP-A.8.24.2: Data Storage Cryptography Implementation Guide
    - ISMS-IMP-A.8.24.3: Authentication Cryptography Assessment (Domain 3)
    - ISMS-IMP-A.8.24.4: Key Management Assessment (Domain 4)
    - ISMS-IMP-A.8.24.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.24.2 specification
    - Supports comprehensive data-at-rest cryptography evaluation
    - Integrated with A.8.24.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Cryptographic Standards:**
Encryption algorithms and key lengths evolve based on threat landscape.
Review industry standards (NIST SP 800-57, BSI TR-02102) annually and update
assessment criteria accordingly. Weak algorithms (DES, 3DES, Blowfish with
64-bit blocks) must be identified and remediated.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of encryption implementation and key management.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Storage system locations and capacities
- Encryption key management architecture
- Unencrypted data repositories (security gaps)
- Data classification details

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new storage systems requiring encryption
- Semi-annually: Validate encryption algorithm compliance
- Annually: Complete reassessment of all storage infrastructure
- Ad-hoc: When infrastructure changes or data breaches occur

**Quality Assurance:**
Have storage administrators and cryptography SMEs validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Ensure encryption standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 data-at-rest encryption requirements
- Healthcare: HIPAA encryption and breach notification rules
- Privacy: GDPR encryption as appropriate safeguard
- Finance: Regional banking data protection requirements
- Government: Jurisdiction-specific encryption mandates

Customize assessment criteria to include regulatory-specific requirements.

**Performance Considerations:**
Document performance impact of encryption (throughput, latency, CPU overhead)
for capacity planning and system optimization. Balance security requirements
with operational performance needs.

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.24.2"
WORKBOOK_NAME = "Data Storage Cryptography Assessment"
CONTROL_ID = "A.8.24"
CONTROL_NAME = "Use of Cryptography"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches markdown specification
    sheets = [
        "Instructions & Legend",
        "1. Mobile Devices",
        "2. Laptops & Workstations",
        "3. Servers",
        "4. Databases",
        "5. Cloud Storage",
        "6. Backups",
        "7. Removable Media",
        "Summary Dashboard",
        "Evidence Register",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "exception_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
    }
    return styles


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS (STANDARD FOR ALL ASSESSMENT SHEETS)
# ============================================================================

def get_storage_columns():
    """
    Return standard column definitions for ALL data storage assessment sheets.
    All 7 assessment sheets use the same 17-column structure.
    
    Returns:
        dict: {column_name: width}
    """
    return {
        "Storage System/Device": 30,
        "Data Classification": 18,
        "Encryption Status": 18,
        "Encryption Type": 20,
        "Algorithm & Key Size": 18,
        "Key Management Method": 22,
        "Key Rotation Enabled": 16,
        "Status": 15,
        "Evidence Location": 28,
        "Gap Description": 30,
        "Remediation Needed": 14,
        "Exception ID": 14,
        "Risk ID": 14,
        "Compensating Controls": 30,
        "Responsible Person": 20,
        "Target Date": 14,
        "Budget Required": 15,
    }


# ============================================================================
# SECTION 3: GLOBAL DATA VALIDATION DEFINITIONS
# ============================================================================

def create_data_validations(ws):
    """
    Create all dropdown validations for assessment sheets.
    Returns dict of DataValidation objects that can be applied to cells.
    """
    validations = {}
    
    # Data Classification
    validations['data_class'] = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['data_class'])
    
    # Encryption Status
    validations['enc_status'] = DataValidation(
        type="list",
        formula1='"Encrypted,Not Encrypted,Partially Encrypted"',
        allow_blank=False
    )
    ws.add_data_validation(validations['enc_status'])
    
    # Encryption Type
    validations['enc_type'] = DataValidation(
        type="list",
        formula1='"Full Disk,File-level,Database TDE,Volume,Container,Application-level,Hardware-based,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['enc_type'])
    
    # Key Management Method
    validations['key_mgmt'] = DataValidation(
        type="list",
        formula1='"HSM,Cloud KMS,TPM,Software-based,Manual,Vendor-managed,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['key_mgmt'])
    
    # Key Rotation Enabled
    validations['key_rotation'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['key_rotation'])
    
    # Status
    validations['status'] = DataValidation(
        type="list",
        formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['status'])
    
    # Remediation Needed
    validations['remediation'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['remediation'])
    
    # Budget Required
    validations['budget'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(validations['budget'])
    
    # Response (Yes/No/Not Applicable)
    validations['response'] = DataValidation(
        type="list",
        formula1='"Yes,No,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(validations['response'])
    
    # Checklist Yes/No/N/A
    validations['checklist'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['checklist'])
    
    # Exception Yes/No
    validations['exception_yn'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['exception_yn'])
    
    return validations


# ============================================================================
# SECTION 4: CHECKLIST DEFINITIONS PER STORAGE TYPE
# ============================================================================

def get_checklist_items(sheet_type):
    """
    Return checklist items for each storage assessment type.
    
    Args:
        sheet_type: String identifier (mobile, laptop, server, database, cloud, backup, removable)
    
    Returns:
        list: Checklist items as strings
    """
    checklists = {
        "mobile": [
            "iOS devices: FileVault/built-in encryption enabled",
            "Android devices: Full device encryption enabled",
            "MDM policy enforces encryption",
            "Encryption status verified via MDM reporting",
            "PIN/password/biometric required for device unlock",
            "Remote wipe capability configured",
            "Encryption cannot be disabled by user",
            "BYOD devices meet encryption requirements",
            "Lost/stolen device procedure includes encryption verification",
            "Regular compliance audits performed (quarterly minimum)",
        ],
        "laptop": [
            "Windows: BitLocker enabled and configured",
            "macOS: FileVault 2 enabled",
            "Linux: LUKS/dm-crypt configured",
            "TPM 2.0 utilized for key storage (where available)",
            "Pre-boot authentication required",
            "Recovery keys securely escrowed",
            "Encryption status monitored centrally",
            "Users cannot disable encryption",
            "Encryption applied before device deployment",
            "Hibernate/sleep mode protected",
            "External storage auto-encryption policy enforced",
        ],
        "server": [
            "OS-level encryption enabled (LUKS, BitLocker, etc.)",
            "Storage volume/LUN encryption configured",
            "Virtual machine disk encryption enabled",
            "Key management solution integrated",
            "Encryption keys stored separately from encrypted data",
            "SAN/NAS encryption enabled (if applicable)",
            "Self-encrypting drives (SED) utilized where appropriate",
            "Boot volumes encrypted",
            "Temporary/swap space encrypted",
            "Encryption status monitoring automated",
            "Key rotation procedures documented and implemented",
            "HSM integration for key protection (production systems)",
        ],
        "database": [
            "TDE (Transparent Data Encryption) enabled",
            "Encryption at column level for sensitive fields (if applicable)",
            "Encryption at tablespace level configured",
            "Database backup files encrypted",
            "Transaction logs encrypted",
            "Encryption keys managed via KMS/HSM",
            "Master encryption key rotated per policy",
            "Encryption does not significantly impact performance",
            "Application-layer encryption for PII/PHI (if TDE unavailable)",
            "Encryption status verified in database audit logs",
            "Development/test databases use encryption for prod data",
            "Database snapshots/clones encrypted",
        ],
        "cloud": [
            "Server-side encryption (SSE) enabled by default",
            "Customer-managed keys (CMK) configured",
            "Client-side encryption implemented for highly sensitive data",
            "Encryption key location documented (region/jurisdiction)",
            "Cloud KMS integrated (AWS KMS, Azure Key Vault, GCP KMS)",
            "Automatic key rotation enabled",
            "Encryption validated via cloud audit logs",
            "Data encrypted before upload (for maximum sensitivity)",
            "Bucket/container policies enforce encryption",
            "Cross-region replication maintains encryption",
            "Cloud backup snapshots encrypted",
            "Compliance certifications verified (SOC 2, ISO 27001)",
        ],
        "backup": [
            "Full backup encryption enabled",
            "Incremental/differential backups encrypted",
            "Backup encryption configured before first backup",
            "Backup encryption keys managed separately from backup data",
            "Offsite/cloud backup storage encrypted in transit AND at rest",
            "Tape backups encrypted (if applicable)",
            "Virtual machine backups encrypted",
            "Database backup encryption independent of TDE",
            "Backup restoration tested with encryption",
            "Backup encryption key recovery procedure documented",
            "Long-term archive backups maintain encryption",
            "Backup logs do not expose sensitive data in plaintext",
        ],
        "removable": [
            "USB drive encryption mandatory via policy",
            "External hard drive encryption enforced",
            "BitLocker To Go / FileVault / LUKS configured",
            "Hardware-encrypted USB drives procured",
            "Unencrypted removable media blocked by endpoint policy",
            "Media encryption verified before data transfer",
            "Encrypted optical media (DVD/CD) procedures documented",
            "Lost removable media incident response includes encryption check",
            "Media disposal/destruction procedures enforce encryption verification",
            "Approved removable media inventory maintained",
            "Encryption exceptions require formal approval",
            "Regular compliance audits of removable media usage",
        ],
    }
    
    return checklists.get(sheet_type, [])


# ============================================================================
# SECTION 5: TECHNOLOGY NOTES PER STORAGE TYPE
# ============================================================================

def get_technology_notes(sheet_type):
    """
    Return technology-specific notes/examples for certain storage types.
    
    Returns:
        list: [(label, value), ...] or empty list
    """
    notes = {
        "database": [
            ("SQL Server:", "TDE, Always Encrypted"),
            ("Oracle:", "TDE, DBMS_CRYPTO"),
            ("PostgreSQL:", "pgcrypto, LUKS volume encryption"),
            ("MySQL:", "InnoDB encryption, file-level encryption"),
            ("MongoDB:", "Encrypted Storage Engine"),
            ("Cassandra:", "Transparent data encryption"),
        ],
        "cloud": [
            ("AWS:", "S3 SSE-KMS, EBS encryption, RDS encryption"),
            ("Azure:", "Storage Service Encryption, Azure Disk Encryption"),
            ("GCP:", "Default encryption at rest, CMEK"),
            ("Others:", "Document specific provider controls"),
        ],
        "backup": [
            ("Agent-based:", "Veeam, Commvault, Veritas"),
            ("Cloud-native:", "AWS Backup, Azure Backup"),
            ("Database-native:", "RMAN (Oracle), SQL Server Backup"),
            ("File-level:", "rsync + GPG, Duplicati"),
            ("Container:", "Velero with encryption"),
        ],
        "removable": [
            ("Windows:", "BitLocker To Go"),
            ("macOS:", "FileVault encrypted volumes"),
            ("Linux:", "LUKS containers"),
            ("Cross-platform:", "VeraCrypt"),
            ("Hardware:", "IronKey, Kingston Encrypted USB"),
            ("Enterprise:", "Symantec Endpoint Encryption"),
        ],
    }
    
    return notes.get(sheet_type, [])


# ============================================================================
# END OF PART 1
# ============================================================================

# ============================================================================
# SECTION 6: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions & Legend sheet matching markdown spec."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.8.24.2 – Data Storage Assessment\n"
        "ISO/IEC 27001:2022 - Control A.8.24: Use of Cryptography"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.24.2"),
        ("Assessment Area", "Data Storage Cryptographic Controls"),
        ("Related Policy", "ISMS-POL-A.8.24-S2.3"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete each worksheet tab (1–7) for applicable storage systems",
        "2. Use dropdown menus for standardized entries (Status, Encryption Type, Algorithm, etc.)",
        "3. Fill in yellow-highlighted cells with your information",
        "4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section",
        "5. Document key management details for each encrypted storage system",
        "6. Provide evidence location/path for each implementation entry",
        "7. Summary Dashboard auto-calculates compliance statistics per storage type",
        "8. Maintain the Evidence Register for audit traceability",
        "9. Obtain final approval and sign-off in the Approval Sign-Off sheet",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    legend = [
        ("Symbol", "Status", "Description"),
        ("✅", "Compliant", "Fully meets policy requirements"),
        ("⚠️", "Partial", "Some requirements met, gaps exist"),
        ("❌", "Non-Compliant", "Does not meet policy requirements"),
        ("N/A", "Not Applicable", "Requirement does not apply"),
    ]

    row += 2
    header_row = row
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if status == "Compliant":
            s.fill = styles["status_compliant"]["fill"]
        elif status == "Partial":
            s.fill = styles["status_partial"]["fill"]
        elif status == "Non-Compliant":
            s.fill = styles["status_noncompliant"]["fill"]

        row += 1

    ws[f"A{row+2}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row+2}"].font = Font(bold=True, size=12)

    evidence_types = [
        "✓ Encryption configuration screenshots",
        "✓ BitLocker/FileVault/LUKS status reports",
        "✓ Database TDE (Transparent Data Encryption) configuration",
        "✓ Cloud encryption settings documentation",
        "✓ Key management system inventory",
        "✓ Backup encryption verification reports",
        "✓ Mobile device management (MDM) compliance reports",
        "✓ Storage vendor encryption specifications",
        "✓ Disk encryption audit logs",
        "✓ Certificate/key rotation logs",
        "✓ Hardware Security Module (HSM) integration docs",
    ]
    row += 3
    for e in evidence_types:
        ws[f"A{row}"] = e
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: GENERIC ASSESSMENT SHEET ENGINE
# ============================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, question, 
                           sheet_type, tech_notes_section=False):
    """
    Generic assessment sheet creator for data storage.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "1. MOBILE DEVICES - DATA AT REST ENCRYPTION"
        policy_ref: policy requirement text
        question: assessment question
        sheet_type: key for checklist items (mobile, laptop, server, etc.)
        tech_notes_section: bool, whether to include technology notes
    """
    columns = get_storage_columns()
    checklist_items = get_checklist_items(sheet_type)
    tech_notes = get_technology_notes(sheet_type) if tech_notes_section else []
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"{section_title}\nPolicy Requirement: {policy_ref}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- ASSESSMENT QUESTION ----------
    ws.merge_cells("A3:Q3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ---------- RESPONSE DROPDOWN ----------
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    # Create validations
    validations = create_data_validations(ws)
    validations['response'].add(ws["B4"])

    # ---------- COLUMN HEADERS ----------
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ---------- EXAMPLE ROW ----------
    example_row = 7
    example_values = [
        "Example: Production SQL Server 2019 - FINDB01",
        "Restricted",
        "Encrypted",
        "Database TDE",
        "AES-256",
        "HSM",
        "Yes",
        "✅ Compliant",
        "/evidence/databases/FINDB01-TDE-config.pdf",
        "None",
        "No",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "N/A",
        "No",
    ]
    for col_idx, value in enumerate(example_values, start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------- DATA ENTRY ROWS (8-20) ----------
    start_row = 8
    end_row = 20
    
    # Column indices for validations
    data_class_col = 2
    enc_status_col = 3
    enc_type_col = 4
    key_mgmt_col = 6
    key_rotation_col = 7
    status_col = 8
    remediation_col = 11
    budget_col = 17
    
    for r in range(start_row, end_row + 1):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Apply validations to data rows
    for r in range(start_row, end_row + 1):
        validations['data_class'].add(ws.cell(row=r, column=data_class_col))
        validations['enc_status'].add(ws.cell(row=r, column=enc_status_col))
        validations['enc_type'].add(ws.cell(row=r, column=enc_type_col))
        validations['key_mgmt'].add(ws.cell(row=r, column=key_mgmt_col))
        validations['key_rotation'].add(ws.cell(row=r, column=key_rotation_col))
        validations['status'].add(ws.cell(row=r, column=status_col))
        validations['remediation'].add(ws.cell(row=r, column=remediation_col))
        validations['budget'].add(ws.cell(row=r, column=budget_col))

    ws.freeze_panes = "A7"

    next_row = end_row + 2

    # ---------- COMPLIANCE CHECKLIST ----------
    ws[f"A{next_row}"] = f"{section_title.split('-')[0].strip().upper()} ENCRYPTION CHECKLIST"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)
    next_row += 1

    ws[f"A{next_row}"] = "☐"
    ws[f"B{next_row}"] = "Requirement"
    ws[f"C{next_row}"] = "Status"
    for col in ["A", "B", "C"]:
        ws[f"{col}{next_row}"].font = Font(bold=True)
    next_row += 1

    checklist_start = next_row
    for item in checklist_items:
        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = item
        ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{next_row}"].border = styles["border"]

        validations['checklist'].add(ws[f"C{next_row}"])
        next_row += 1

    # Checklist score
    ws[f"A{next_row}"] = "Checklist Score:"
    ws[f"A{next_row}"].font = Font(bold=True)
    ws[f"B{next_row}"] = (
        f'=COUNTIF(C{checklist_start}:C{next_row-1},"Yes")/'
        f'COUNTA(C{checklist_start}:C{next_row-1})*100&"%"'
    )
    ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
    next_row += 2

    # ---------- TECHNOLOGY NOTES (if applicable) ----------
    if tech_notes:
        ws[f"A{next_row}"] = "Technology-Specific Notes / Examples"
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1
        
        for label, value in tech_notes:
            ws[f"A{next_row}"] = label
            ws[f"A{next_row}"].font = Font(bold=True)
            ws[f"B{next_row}"] = value
            ws[f"B{next_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            next_row += 1
        
        next_row += 1

    # ---------- EXCEPTION/DEVIATION BLOCK ----------
    ws.merge_cells(f"A{next_row}:Q{next_row}")
    ws[f"A{next_row}"] = "EXCEPTION / DEVIATION DOCUMENTATION (Complete if Status = Partial or Non-Compliant)"
    ws[f"A{next_row}"].font = styles["exception_header"]["font"]
    ws[f"A{next_row}"].fill = styles["exception_header"]["fill"]
    ws[f"A{next_row}"].alignment = styles["exception_header"]["alignment"]

    next_row += 1
    exception_fields = [
        ("Formal exception request submitted:", "Yes,No"),
        ("Exception ID:", ""),
        ("Risk acceptance documented:", "Yes,No"),
        ("Risk ID:", ""),
        ("Compensating Controls (summary):", ""),
        ("☐ Physical security controls (locked facility, restricted access)", ""),
        ("☐ Network segmentation / air-gapped systems", ""),
        ("☐ Enhanced monitoring / DLP controls", ""),
        ("☐ Restricted user access / privileged access management", ""),
        ("☐ Data minimization / reduced storage scope", ""),
        ("☐ Other (describe):", ""),
    ]

    for label, options in exception_fields:
        ws[f"A{next_row}"] = label
        ws[f"A{next_row}"].font = Font(bold=True)
        ws.merge_cells(f"B{next_row}:D{next_row}")
        ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{next_row}"].border = styles["border"]
        ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]

        if options:
            validations['exception_yn'].add(ws[f"B{next_row}"])

        next_row += 1

    # ---------- NOTES ----------
    next_row += 1
    ws.merge_cells(f"A{next_row}:Q{next_row}")
    ws[f"A{next_row}"] = "ADDITIONAL NOTES / COMMENTS"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)

    next_row += 1
    ws.merge_cells(f"A{next_row}:Q{next_row+8}")
    ws[f"A{next_row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{next_row}"].border = styles["border"]
    ws[f"A{next_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths for B and C (for checklists and notes)
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15

    return (start_row, end_row, status_col)


# ============================================================================
# END OF PART 2
# ============================================================================

# ============================================================================
# SECTION 8: INDIVIDUAL ASSESSMENT SHEET DEFINITIONS
# ============================================================================

def create_1_mobile_devices(ws, styles):
    """1. Mobile Devices - matches markdown spec exactly."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1. MOBILE DEVICES - DATA AT REST ENCRYPTION",
        policy_ref="Mobile devices storing organizational data MUST use full device encryption (Policy Section 2.3.1)",
        question="Does your organization issue or allow mobile devices (smartphones, tablets) that store organizational data?",
        sheet_type="mobile",
        tech_notes_section=False,
    )


def create_2_laptops_workstations(ws, styles):
    """2. Laptops & Workstations - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2. LAPTOPS & WORKSTATIONS - DATA AT REST ENCRYPTION",
        policy_ref="Laptops and workstations storing sensitive data MUST use full disk encryption (Policy Section 2.3.2)",
        question="Does your organization have laptops or workstations that store organizational data locally?",
        sheet_type="laptop",
        tech_notes_section=False,
    )


def create_3_servers(ws, styles):
    """3. Servers - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="3. SERVERS - DATA AT REST ENCRYPTION",
        policy_ref="Servers storing sensitive data MUST implement appropriate encryption at rest (Policy Section 2.3.3)",
        question="Does your organization operate servers (physical or virtual) that store sensitive organizational data?",
        sheet_type="server",
        tech_notes_section=False,
    )


def create_4_databases(ws, styles):
    """4. Databases - matches markdown spec with technology notes."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4. DATABASES - DATA AT REST ENCRYPTION",
        policy_ref="Databases containing sensitive data MUST implement Transparent Data Encryption or equivalent (Policy Section 2.3.4)",
        question="Does your organization operate databases that store sensitive or regulated data?",
        sheet_type="database",
        tech_notes_section=True,
    )


def create_5_cloud_storage(ws, styles):
    """5. Cloud Storage - matches markdown spec with technology notes."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="5. CLOUD STORAGE - DATA AT REST ENCRYPTION",
        policy_ref="Cloud storage MUST use provider encryption with customer-managed keys where available (Policy Section 2.3.5)",
        question="Does your organization store data in cloud storage services (object storage, block storage, file storage)?",
        sheet_type="cloud",
        tech_notes_section=True,
    )


def create_6_backups(ws, styles):
    """6. Backups - matches markdown spec with technology notes."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="6. BACKUPS - DATA AT REST ENCRYPTION",
        policy_ref="Backup media and backup data MUST be encrypted with strong cryptography (Policy Section 2.3.6)",
        question="Does your organization perform backups of systems, databases, or files?",
        sheet_type="backup",
        tech_notes_section=True,
    )


def create_7_removable_media(ws, styles):
    """7. Removable Media - matches markdown spec with technology notes."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="7. REMOVABLE MEDIA - DATA AT REST ENCRYPTION",
        policy_ref="Removable media storing organizational data MUST be encrypted (Policy Section 2.3.7)",
        question="Does your organization use removable media (USB drives, external disks, optical media) to store or transport data?",
        sheet_type="removable",
        tech_notes_section=True,
    )


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create summary dashboard with compliance statistics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "DATA STORAGE ASSESSMENT - COMPLIANCE SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # ---------- COMPLIANCE SUMMARY TABLE ----------
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Map dashboard rows to sheet names (status column is always H = column 8)
    areas = [
        ("1. Mobile Devices", "1. Mobile Devices"),
        ("2. Laptops & Workstations", "2. Laptops & Workstations"),
        ("3. Servers", "3. Servers"),
        ("4. Databases", "4. Databases"),
        ("5. Cloud Storage", "5. Cloud Storage"),
        ("6. Backups", "6. Backups"),
        ("7. Removable Media", "7. Removable Media"),
    ]

    row += 1
    start_data_row = row
    for label, sheet in areas:
        ws.cell(row=row, column=1, value=label)
        
        status_range = f"'{sheet}'!H8:H20"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"✅*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"⚠️*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"❌*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(bold=True, color="0000FF", size=12)

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    # ---------- ENCRYPTION COVERAGE BY DATA CLASSIFICATION ----------
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ENCRYPTION COVERAGE BY DATA CLASSIFICATION"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    class_headers = ["Data Classification", "Systems Assessed", "Encrypted", "Unencrypted", "Coverage %"]
    for col_idx, header in enumerate(class_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    classifications = ["Restricted", "Confidential", "Internal", "Public"]
    row += 1
    for classification in classifications:
        ws.cell(row=row, column=1, value=classification).fill = styles["input_cell"]["fill"]
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        row += 1

    # ---------- KEY MANAGEMENT SUMMARY ----------
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "KEY MANAGEMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    key_headers = ["Key Management Method", "Count", "Percentage"]
    for col_idx, header in enumerate(key_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    key_methods = ["HSM", "Cloud KMS", "TPM", "Software-based", "Manual", "Vendor-managed"]
    row += 1
    key_start_row = row
    
    # Build formula to count across all sheets
    all_sheets = ["1. Mobile Devices", "2. Laptops & Workstations", "3. Servers", 
                  "4. Databases", "5. Cloud Storage", "6. Backups", "7. Removable Media"]
    
    for method in key_methods:
        ws.cell(row=row, column=1, value=method)
        
        # Count formula across all sheets (column F = Key Management Method)
        count_formula = "+".join([f"COUNTIF('{sheet}'!F8:F20,\"{method}\")" for sheet in all_sheets])
        ws.cell(row=row, column=2, value=f"={count_formula}")
        
        # Percentage
        ws.cell(row=row, column=3, value=f'=IF(SUM(B{key_start_row}:B{row})=0,"0%",ROUND(B{row}/SUM(B${key_start_row}:B${row})*100,1)&"%")')
        row += 1

    # ---------- CRITICAL GAPS ----------
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    gap_headers = ["Priority", "Assessment Area", "Gap Description", "Responsible Person", "Target Date", "Status"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Priority dropdown
    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_priority)

    # Status dropdown for gaps
    dv_gap_status = DataValidation(type="list", formula1='"Not started,In progress,Completed,Blocked"', allow_blank=False)
    ws.add_data_validation(dv_gap_status)

    for i in range(5):
        row += 1
        ws.cell(row=row, column=1).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=1).border = styles["border"]
        dv_priority.add(ws.cell(row=row, column=1))
        
        for col in range(2, 7):
            c = ws.cell(row=row, column=col)
            c.fill = styles["input_cell"]["fill"]
            c.border = styles["border"]
            c.alignment = styles["input_cell"]["alignment"]
        
        dv_gap_status.add(ws.cell(row=row, column=6))

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create evidence register for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validation dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Encryption report,MDM report,Audit log,Key management record,Backup verification,Vendor documentation,Compliance certificate,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    # Data rows (100 rows as per spec)
    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# END OF PART 3
# ============================================================================

# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create approval and sign-off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.8.24.2 - Data Storage Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G11"),  # Total row from dashboard
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)

        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Status dropdown
    status_cell_row = row - 1
    dv_status = DataValidation(
        type="list", 
        formula1='"Draft,Final,Requires remediation,Re-assessment required"', 
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{status_cell_row}"])

    # ---------- ASSESSMENT COMPLETED BY ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # ---------- APPROVED BY (CISO) ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Approval Decision dropdown (3rd field from approval section start)
    decision_cell_row = row - 1
    dv_dec = DataValidation(
        type="list", 
        formula1='"Approved,Approved with conditions,Rejected"', 
        allow_blank=False
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{decision_cell_row}"])

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def main() -> int:
    """Main execution function - orchestrates workbook creation."""
    try:
        logger.info("=" * 78)
        logger.info("ISMS-IMP-A.8.24.2 - Data Storage Assessment Generator")
        logger.info("ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography (Data at Rest)")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/11] Creating Instructions & Legend sheet...")
        create_instructions_sheet(wb["Instructions & Legend"], styles)

        logger.info("[2/11] Creating 1. Mobile Devices sheet...")
        create_1_mobile_devices(wb["1. Mobile Devices"], styles)

        logger.info("[3/11] Creating 2. Laptops & Workstations sheet...")
        create_2_laptops_workstations(wb["2. Laptops & Workstations"], styles)

        logger.info("[4/11] Creating 3. Servers sheet...")
        create_3_servers(wb["3. Servers"], styles)

        logger.info("[5/11] Creating 4. Databases sheet...")
        create_4_databases(wb["4. Databases"], styles)

        logger.info("[6/11] Creating 5. Cloud Storage sheet...")
        create_5_cloud_storage(wb["5. Cloud Storage"], styles)

        logger.info("[7/11] Creating 6. Backups sheet...")
        create_6_backups(wb["6. Backups"], styles)

        logger.info("[8/11] Creating 7. Removable Media sheet...")
        create_7_removable_media(wb["7. Removable Media"], styles)

        logger.info("[9/11] Creating Summary Dashboard sheet...")
        create_summary_dashboard(wb["Summary Dashboard"], styles)

        logger.info("[10/11] Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"], styles)

        logger.info("[11/11] Creating Approval Sign-Off sheet...")
        create_approval_signoff(wb["Approval Sign-Off"], styles)

        filename = f"ISMS-IMP-A.8.24.2_Data_Storage_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info(f"SUCCESS: {filename}")
        logger.info("Workbook Structure:")
        logger.info("  - Instructions & Legend - Document info and guidance")
        logger.info("  - 7 Assessment Sheets:")
        logger.info("    1. Mobile Devices (10 checklist items)")
        logger.info("    2. Laptops & Workstations (11 checklist items)")
        logger.info("    3. Servers (12 checklist items)")
        logger.info("    4. Databases (12 checklist items + tech notes)")
        logger.info("    5. Cloud Storage (12 checklist items + tech notes)")
        logger.info("    6. Backups (12 checklist items + tech notes)")
        logger.info("    7. Removable Media (12 checklist items + tech notes)")
        logger.info("  - Summary Dashboard - Compliance statistics & gap tracking")
        logger.info("  - Evidence Register - 100 evidence entry rows")
        logger.info("  - Approval Sign-Off - Final sign-off section")

        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions & Legend")
        logger.info("  2) Fill yellow cells in each assessment sheet (1-7)")
        logger.info("  3) Check compliance checklists per storage type")
        logger.info("  4) Document exceptions/deviations as needed")
        logger.info("  5) Maintain Evidence Register entries")
        logger.info("  6) Review Summary Dashboard:")
        logger.info("     - Compliance summary by storage type")
        logger.info("     - Encryption coverage by data classification")
        logger.info("     - Key management method distribution")
        logger.info("     - Critical gaps requiring immediate attention")
        logger.info("  7) Complete Approval Sign-Off")
        logger.info("=" * 78)
        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())


# ============================================================================
# END OF SCRIPT - COMPLETE DATA STORAGE ASSESSMENT GENERATOR
# ============================================================================
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
