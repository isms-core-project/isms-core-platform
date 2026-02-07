#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-IMP-A.8.24.1 - Data Transmission Assessment
Comprehensive Data Population Script for CISO Presentation

Populates ALL sheets with production-quality demo data:
- 12 Assessment sheets (HTTPS, Email, VPN, SSH, RDP, API, DB, Wireless, Cloud)
- Summary Dashboard (auto-calculated)
- Evidence Register (25+ evidence documents)
- Approval Sign-Off (complete workflow)

Usage:
    python3 populate_a824_1_data_transmission.py ISMS-IMP-A.8.24.1.xlsx

Generated: 2026-01-13
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def populate_assessment_sheets(wb):
    """Populate all 12 assessment sheets with realistic data."""

    # Columns: Service/System, Current TLS Version, Certificate Source,
    #          Certificate Validity, Status, Evidence Location, Gap Description, Remediation Needed

    sheet_data = {
        "1.1 External HTTPS-TLS": [
            ["www.company.com", "TLS 1.3", "DigiCert SHA-256", "Valid until 2026-12-31", "✅ Compliant", "SSL Labs Report A+", "", "No"],
            ["portal.company.com", "TLS 1.3", "Let's Encrypt", "Valid until 2026-06-15", "✅ Compliant", "Certbot auto-renew logs", "", "No"],
            ["api.company.com", "TLS 1.3", "DigiCert EV SSL", "Valid until 2027-03-20", "✅ Compliant", "AWS Certificate Manager", "", "No"],
            ["mail.company.com", "TLS 1.2", "GoDaddy SSL", "Expires 2026-02-28", "⚠️ Partial", "Renewal scheduled Q1", "Upgrade to TLS 1.3", "Yes"],
            ["mobile-api.company.com", "TLS 1.3", "DigiCert", "Valid until 2026-11-30", "✅ Compliant", "Auto-rotation enabled", "", "No"],
            ["partners.company.com", "TLS 1.3", "GlobalSign", "Valid until 2027-01-15", "✅ Compliant", "SAN certificate", "", "No"],
            ["cdn.company.com", "TLS 1.3", "Cloudflare Universal", "Valid until 2027-06-01", "✅ Compliant", "Cloudflare managed", "", "No"],
            ["admin.company.com", "TLS 1.3", "DigiCert EV SSL", "Valid until 2026-09-15", "✅ Compliant", "EV cert + HSM backup", "", "No"],
            ["shop.company.com", "TLS 1.3", "Sectigo", "Valid until 2026-08-20", "✅ Compliant", "E-commerce certificate", "", "No"],
            ["docs.company.com", "TLS 1.3", "Let's Encrypt", "Valid until 2026-04-10", "✅ Compliant", "Auto-renew via certbot", "", "No"],
            ["legacy.company.com", "TLS 1.1", "Self-signed", "Expired 2024-01-01", "❌ Non-Compliant", "Legacy system logs", "URGENT: Decommission or upgrade", "Yes"],
            ["vpn.company.com", "TLS 1.2", "Internal CA", "Valid until 2025-12-31", "⚠️ Partial", "Internal PKI", "Plan TLS 1.3 upgrade", "Yes"],
            ["test.company.com", "TLS 1.2", "Self-signed", "Valid until 2026-03-01", "⚠️ Partial", "Test environment", "Use valid certificate", "Yes"],
            ["blog.company.com", "TLS 1.3", "Let's Encrypt", "Valid until 2026-07-15", "✅ Compliant", "WordPress auto-SSL", "", "No"],
            ["status.company.com", "TLS 1.3", "DigiCert", "Valid until 2026-10-01", "✅ Compliant", "Status page service", "", "No"],
        ],

        "1.2 Internal HTTPS-TLS": [
            ["intranet.internal", "TLS 1.3", "Internal CA", "Valid until 2026-12-01", "✅ Compliant", "AD Certificate Services", "", "No"],
            ["gitlab.internal", "TLS 1.3", "Internal CA", "Valid until 2026-11-15", "✅ Compliant", "GitLab SSL config", "", "No"],
            ["jenkins.internal", "TLS 1.3", "Internal CA", "Valid until 2026-10-20", "✅ Compliant", "Jenkins HTTPS setup", "", "No"],
            ["grafana.internal", "TLS 1.3", "Internal CA", "Valid until 2026-09-30", "✅ Compliant", "Grafana SSL logs", "", "No"],
            ["prometheus.internal", "TLS 1.3", "Internal CA", "Valid until 2026-09-15", "✅ Compliant", "Prometheus config", "", "No"],
            ["vault.internal", "TLS 1.3", "Internal CA + HSM", "Valid until 2027-01-01", "✅ Compliant", "Vault audit logs", "", "No"],
            ["elasticsearch.internal", "TLS 1.3", "Internal CA", "Valid until 2026-08-20", "✅ Compliant", "ES cluster config", "", "No"],
            ["kibana.internal", "TLS 1.3", "Internal CA", "Valid until 2026-08-20", "✅ Compliant", "Kibana SSL settings", "", "No"],
            ["confluence.internal", "TLS 1.2", "Internal CA", "Valid until 2025-12-31", "⚠️ Partial", "Confluence config", "Upgrade to TLS 1.3", "Yes"],
            ["jira.internal", "TLS 1.2", "Internal CA", "Valid until 2025-12-31", "⚠️ Partial", "Jira SSL config", "Upgrade to TLS 1.3", "Yes"],
            ["monitoring.internal", "TLS 1.3", "Internal CA", "Valid until 2026-07-01", "✅ Compliant", "Monitoring stack", "", "No"],
            ["backup.internal", "TLS 1.2", "Self-signed", "Expired 2025-11-01", "❌ Non-Compliant", "Backup server logs", "Replace with valid cert", "Yes"],
        ],

        "2.1 Email Encryption": [
            ["Microsoft 365", "TLS 1.3 + S/MIME", "DigiCert", "85% user adoption", "⚠️ Partial", "M365 Admin Center", "Increase S/MIME adoption to 95%", "Yes"],
            ["Gmail Business", "TLS 1.3", "Google Trust Services", "100% enforced", "✅ Compliant", "Google Workspace Admin", "", "No"],
            ["Internal Mail Server", "TLS 1.2 + PGP", "Internal CA", "45% PGP adoption", "⚠️ Partial", "Postfix logs", "Promote PGP key usage", "Yes"],
            ["Executive Email", "TLS 1.3 + S/MIME", "DigiCert", "100% mandatory", "✅ Compliant", "Certificate store audit", "", "No"],
            ["Partner Email Gateway", "TLS 1.2", "StartTLS", "90% TLS coverage", "✅ Compliant", "Mail gateway logs", "", "No"],
            ["Marketing Automation", "TLS 1.3", "SendGrid", "100%", "✅ Compliant", "SendGrid dashboard", "", "No"],
            ["Support Tickets", "TLS 1.3", "Zendesk", "100%", "✅ Compliant", "Zendesk security", "", "No"],
            ["Notification System", "TLS 1.3", "AWS SES", "100%", "✅ Compliant", "SES configuration", "", "No"],
        ],

        "2.2 Digital Signatures": [
            ["Executive Contracts", "RSA-4096 + SHA-256", "DocuSign", "100% signed", "✅ Compliant", "DocuSign audit trail", "", "No"],
            ["Financial Documents", "RSA-4096 + SHA-256", "Adobe Sign", "100% required", "✅ Compliant", "Adobe audit logs", "", "No"],
            ["Code Signing", "RSA-4096 + SHA-256", "YubiKey HSM", "All releases", "✅ Compliant", "CI/CD signing logs", "", "No"],
            ["Internal Approvals", "RSA-2048 + SHA-256", "Internal PKI", "80% adoption", "⚠️ Partial", "Approval workflow logs", "Increase to 95%", "Yes"],
            ["PDF Documents", "RSA-4096 + SHA-256", "Adobe Acrobat", "All sensitive docs", "✅ Compliant", "Document management", "", "No"],
            ["Email Signatures", "RSA-2048 + SHA-256", "S/MIME", "Executive only", "⚠️ Partial", "Certificate logs", "Expand to all users", "Yes"],
            ["Legal Documents", "RSA-4096 + SHA-256", "DocuSign + PKI", "100% required", "✅ Compliant", "Legal system logs", "", "No"],
        ],

        "3.1 File Transfer Protocols": [
            ["SFTP Production", "SSH-2 (RSA-4096)", "Key-based", "Port 22", "✅ Compliant", "SFTP server logs", "", "No"],
            ["FTPS Partner Gateway", "TLS 1.3", "Certificate-based", "Port 990", "✅ Compliant", "FileZilla logs", "", "No"],
            ["SCP Internal", "SSH-2 (Ed25519)", "Key-based", "Port 22", "✅ Compliant", "SSH audit logs", "", "No"],
            ["Legacy FTP", "FTP (plaintext)", "Password", "Port 21", "❌ Non-Compliant", "Old FTP logs", "URGENT: Migrate to SFTP", "Yes"],
            ["HTTPS File Upload", "TLS 1.3", "Certificate", "Port 443", "✅ Compliant", "Web server logs", "", "No"],
            ["rsync Backups", "SSH-2 (RSA-4096)", "Key-based", "Port 22", "✅ Compliant", "Backup logs", "", "No"],
            ["AWS S3 Transfer", "TLS 1.3", "IAM + KMS", "Port 443", "✅ Compliant", "CloudTrail logs", "", "No"],
            ["Azure Files", "SMB 3.1.1 + TLS 1.3", "Kerberos + cert", "Port 445", "✅ Compliant", "Azure Storage logs", "", "No"],
        ],

        "4.1 VPN": [
            ["Employee VPN (OpenVPN)", "TLS 1.3", "AES-256-GCM + RSA-4096", "MFA required", "✅ Compliant", "VPN server logs + MFA logs", "", "No"],
            ["Site-to-Site VPN (IPSec)", "IKEv2", "AES-256-CBC + DH Group 14", "Pre-shared key", "⚠️ Partial", "Firewall logs", "Add certificate auth + DH 20", "Yes"],
            ["Executive VPN (WireGuard)", "WireGuard", "ChaCha20-Poly1305 + Curve25519", "MFA + device cert", "✅ Compliant", "WireGuard logs + MDM", "", "No"],
            ["Partner VPN (SSL)", "TLS 1.3", "AES-256-GCM + ECDHE", "Certificate + MFA", "✅ Compliant", "SSL VPN access logs", "", "No"],
            ["Contractor VPN (OpenVPN)", "TLS 1.2", "AES-128-CBC + RSA-2048", "Password only", "❌ Non-Compliant", "Legacy VPN config", "Upgrade: TLS 1.3 + MFA + AES-256", "Yes"],
            ["Remote Admin VPN", "TLS 1.3", "AES-256-GCM + RSA-4096", "MFA + FIDO2", "✅ Compliant", "Admin VPN audit logs", "", "No"],
            ["Vendor VPN (Temporary)", "TLS 1.3", "AES-256-GCM", "Time-limited certs", "✅ Compliant", "Vendor access logs", "", "No"],
        ],

        "4.2 SSH": [
            ["Production Servers", "SSH-2 (Ed25519)", "Key-based + FIDO2", "Bastion required", "✅ Compliant", "SSH audit logs + Bastion", "", "No"],
            ["Development Servers", "SSH-2 (RSA-4096)", "Key-based", "Direct access", "✅ Compliant", "Dev SSH logs", "", "No"],
            ["Database Servers", "SSH-2 (Ed25519)", "Key + MFA", "Bastion + audit", "✅ Compliant", "DB SSH logs", "", "No"],
            ["Network Devices", "SSH-2 (RSA-2048)", "Password + TACACS+", "Cisco devices", "⚠️ Partial", "TACACS+ logs", "Migrate to key-based", "Yes"],
            ["IoT Devices", "SSH-2 (RSA-2048)", "Default passwords", "Embedded systems", "❌ Non-Compliant", "IoT inventory", "Implement key management", "Yes"],
            ["Jump/Bastion Host", "SSH-2 (Ed25519)", "Key + FIDO2 + session recording", "All access logged", "✅ Compliant", "Bastion audit logs", "", "No"],
            ["Git Repository Access", "SSH-2 (Ed25519)", "Key-based", "GitLab/GitHub", "✅ Compliant", "Git SSH logs", "", "No"],
            ["Backup Server Access", "SSH-2 (RSA-4096)", "Key-based", "Restricted IPs", "✅ Compliant", "Backup SSH logs", "", "No"],
        ],

        "4.3 RDP": [
            ["Windows Servers", "RDP with NLA + TLS 1.3", "AES-256", "MFA enforced", "✅ Compliant", "RDP logs + MFA events", "", "No"],
            ["Windows Workstations", "RDP with NLA + TLS 1.2", "AES-128", "Password only", "⚠️ Partial", "Workstation RDP logs", "Enforce MFA + TLS 1.3", "Yes"],
            ["Terminal Servers", "RDP with NLA + TLS 1.3", "AES-256", "MFA + smartcard", "✅ Compliant", "Terminal server logs", "", "No"],
            ["Legacy Systems", "RDP without NLA", "RC4 (weak)", "Password", "❌ Non-Compliant", "Legacy RDP logs", "URGENT: Enable NLA + TLS 1.3", "Yes"],
            ["Admin Workstations", "RDP with NLA + TLS 1.3", "AES-256-GCM", "FIDO2 required", "✅ Compliant", "Admin RDP audit", "", "No"],
            ["Jump Server RDP", "RDP with NLA + TLS 1.3", "AES-256 + session recording", "Certificate + MFA", "✅ Compliant", "Jump server logs", "", "No"],
        ],

        "5.1 API Security": [
            ["REST API (Public)", "TLS 1.3", "AES-256-GCM + ECDHE", "JWT + OAuth 2.0", "✅ Compliant", "API Gateway logs", "", "No"],
            ["GraphQL API", "TLS 1.3", "AES-256-GCM", "API Key + JWT", "✅ Compliant", "GraphQL server logs", "", "No"],
            ["Internal Microservices", "mTLS (TLS 1.3)", "AES-256-GCM", "Service mesh certs", "✅ Compliant", "Istio/Linkerd logs", "", "No"],
            ["Partner API", "TLS 1.3", "AES-256-GCM", "Client certificates", "✅ Compliant", "Partner API logs", "", "No"],
            ["Mobile App API", "TLS 1.3 + Certificate Pinning", "AES-256-GCM", "JWT + refresh tokens", "✅ Compliant", "Mobile API logs", "", "No"],
            ["Legacy SOAP API", "TLS 1.2", "AES-128-CBC", "WS-Security (username)", "⚠️ Partial", "SOAP logs", "Upgrade TLS 1.3 + OAuth", "Yes"],
            ["Webhook Endpoints", "TLS 1.3", "AES-256-GCM", "HMAC signatures", "✅ Compliant", "Webhook logs", "", "No"],
            ["Admin API", "TLS 1.3 + mTLS", "AES-256-GCM", "Client cert + RBAC", "✅ Compliant", "Admin API audit", "", "No"],
        ],

        "6.1 Database Connections": [
            ["PostgreSQL Production", "TLS 1.3 + SCRAM-SHA-256", "AES-256-GCM", "Certificate required", "✅ Compliant", "PostgreSQL logs", "", "No"],
            ["MySQL Cluster", "TLS 1.3", "AES-256", "Certificate + password", "✅ Compliant", "MySQL SSL status", "", "No"],
            ["SQL Server", "TLS 1.3 + Kerberos", "AES-256", "Windows Auth", "✅ Compliant", "SQL Server DMVs", "", "No"],
            ["MongoDB", "TLS 1.3 + SCRAM", "AES-256-GCM", "X.509 certificates", "✅ Compliant", "MongoDB logs", "", "No"],
            ["Redis", "TLS 1.3", "AES-256-GCM", "ACL + password", "✅ Compliant", "Redis ACL logs", "", "No"],
            ["Oracle Database", "TLS 1.2 + Oracle Native", "AES-192", "Password only", "⚠️ Partial", "Oracle audit", "Upgrade TLS 1.3 + AES-256", "Yes"],
            ["Elasticsearch", "TLS 1.3 + SAML", "AES-256-GCM", "Certificate + RBAC", "✅ Compliant", "ES security logs", "", "No"],
            ["Cassandra", "TLS 1.3", "AES-256", "Certificate auth", "✅ Compliant", "Cassandra audit", "", "No"],
            ["MariaDB", "TLS 1.3", "AES-256", "Certificate required", "✅ Compliant", "MariaDB SSL logs", "", "No"],
            ["DynamoDB", "TLS 1.3 + IAM", "AES-256-GCM", "IAM roles", "✅ Compliant", "CloudTrail logs", "", "No"],
        ],

        "6.2 Wireless Networks": [
            ["Corporate WiFi", "WPA3-Enterprise", "AES-256 + 802.1X", "RADIUS + certificates", "✅ Compliant", "RADIUS logs + AP logs", "", "No"],
            ["Guest WiFi", "WPA3-Personal", "AES-256", "Captive portal", "✅ Compliant", "Guest access logs", "", "No"],
            ["IoT WiFi", "WPA2-Enterprise", "AES-256 + 802.1X", "Device certificates", "⚠️ Partial", "IoT VLAN logs", "Migrate to WPA3", "Yes"],
            ["Executive Floor", "WPA3-Enterprise", "AES-256 + 802.1X", "RADIUS + FIDO2", "✅ Compliant", "Executive AP logs", "", "No"],
            ["Conference Rooms", "WPA3-Personal", "AES-256", "Daily rotating PSK", "✅ Compliant", "Conference WiFi logs", "", "No"],
            ["Warehouse WiFi", "WPA2-Personal", "AES-128", "Shared PSK", "❌ Non-Compliant", "Warehouse AP logs", "Upgrade WPA3 + 802.1X", "Yes"],
            ["Data Center WiFi", "Disabled", "N/A", "Wired only", "✅ Compliant", "DC access policy", "", "No"],
        ],

        "7.1 Cloud Transmission": [
            ["AWS Services", "TLS 1.3", "AES-256-GCM", "IAM + KMS", "✅ Compliant", "CloudTrail + VPC Flow", "", "No"],
            ["Azure Services", "TLS 1.3", "AES-256-GCM", "Azure AD + Key Vault", "✅ Compliant", "Azure Monitor", "", "No"],
            ["GCP Services", "TLS 1.3", "AES-256-GCM", "IAM + Cloud KMS", "✅ Compliant", "Cloud Audit Logs", "", "No"],
            ["AWS VPC Peering", "Encrypted in-transit", "AES-256-GCM", "VPC routing", "✅ Compliant", "VPC peering logs", "", "No"],
            ["Azure ExpressRoute", "Private connection + IPsec", "AES-256", "MACsec enabled", "✅ Compliant", "ExpressRoute metrics", "", "No"],
            ["GCP Interconnect", "Private connection + IPsec", "AES-256", "Cloud VPN", "✅ Compliant", "Interconnect logs", "", "No"],
            ["Multi-cloud VPN", "TLS 1.3 + IPsec", "AES-256-GCM", "Site-to-site", "✅ Compliant", "Multi-cloud logs", "", "No"],
            ["CDN (CloudFront)", "TLS 1.3", "AES-256-GCM", "AWS managed", "✅ Compliant", "CloudFront logs", "", "No"],
            ["CDN (Azure CDN)", "TLS 1.3", "AES-256-GCM", "Azure managed", "✅ Compliant", "CDN analytics", "", "No"],
            ["SaaS Applications", "TLS 1.3", "AES-256", "OAuth + SAML", "✅ Compliant", "SSO audit logs", "", "No"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    %s: %d entries", sheet_name, count)

    return total_written


def populate_evidence_register(wb):
    """Populate Evidence Register with realistic evidence documents."""

    if "Evidence Register" not in wb.sheetnames:
        return 0

    ws = wb["Evidence Register"]

    # Columns (must match generator): Evidence ID, Assessment Area, Evidence Type,
    #          Description, Location/Path, Date Collected, Collected By, Verification Status
    # Evidence Type dropdown: Configuration file, Screenshot, Network scan, Documentation, Vendor spec, Certificate inventory, Audit log, Compliance report, Other
    # Verification Status dropdown: Verified, Pending verification, Not verified, Requires update

    evidence_data = [
        ["EVD-824.1-001", "1.1 External HTTPS-TLS", "Compliance report", "SSL Labs Security Report - www.company.com", "SharePoint/Evidence/TLS", "2026-01-10", "Security Team", "Verified"],
        ["EVD-824.1-002", "1.1 External HTTPS-TLS", "Certificate inventory", "Certificate Inventory Export", "SharePoint/Evidence/Certificates", "2026-01-10", "Security Team", "Verified"],
        ["EVD-824.1-003", "1.1 External HTTPS-TLS", "Audit log", "Let's Encrypt Auto-Renewal Logs", "S3/evidence/certbot-logs", "2026-01-10", "DevOps", "Verified"],
        ["EVD-824.1-004", "1.1 External HTTPS-TLS", "Screenshot", "AWS ACM Certificate Configuration", "SharePoint/Evidence/AWS", "2026-01-10", "Cloud Team", "Verified"],
        ["EVD-824.1-005", "1.2 Internal HTTPS-TLS", "Documentation", "Internal PKI Certificate Policy", "SharePoint/Policies", "2025-06-01", "Security Team", "Verified"],
        ["EVD-824.1-006", "1.2 Internal HTTPS-TLS", "Compliance report", "AD Certificate Services Audit Report", "SharePoint/Audits/2025", "2025-12-15", "Internal Audit", "Verified"],
        ["EVD-824.1-007", "2.1 Email Encryption", "Compliance report", "M365 Admin Center - S/MIME Adoption Report", "SharePoint/Evidence/M365", "2026-01-09", "IT Admin", "Verified"],
        ["EVD-824.1-008", "2.1 Email Encryption", "Screenshot", "Google Workspace Security Settings", "SharePoint/Evidence/Google", "2026-01-09", "IT Admin", "Verified"],
        ["EVD-824.1-009", "2.1 Email Encryption", "Configuration file", "Postfix TLS Configuration File", "Git/configs/postfix", "2025-11-20", "Email Admin", "Verified"],
        ["EVD-824.1-010", "2.1 Email Encryption", "Compliance report", "Executive S/MIME Certificate Audit", "SharePoint/Audits/Certs", "2025-12-01", "Security Team", "Verified"],
        ["EVD-824.1-011", "2.2 Digital Signatures", "Audit log", "DocuSign Audit Trail Report", "DocuSign/Audit-Trails", "2026-01-08", "Legal Team", "Verified"],
        ["EVD-824.1-012", "2.2 Digital Signatures", "Screenshot", "Adobe Sign Security Configuration", "SharePoint/Evidence/Adobe", "2026-01-08", "Legal Team", "Verified"],
        ["EVD-824.1-013", "2.2 Digital Signatures", "Certificate inventory", "Code Signing Certificate - YubiKey HSM", "PKI/Code-Signing", "2025-09-15", "DevOps", "Verified"],
        ["EVD-824.1-014", "2.2 Digital Signatures", "Audit log", "CI/CD Pipeline Signing Logs", "Jenkins/Logs/Signing", "2026-01-07", "DevOps", "Verified"],
        ["EVD-824.1-015", "3.1 File Transfer", "Configuration file", "SFTP Server Configuration", "Git/configs/sftp", "2025-10-01", "SysAdmin", "Verified"],
        ["EVD-824.1-016", "3.1 File Transfer", "Audit log", "FTPS Partner Gateway Access Logs", "S3/logs/ftps", "2026-01-10", "Network Team", "Verified"],
        ["EVD-824.1-017", "3.1 File Transfer", "Documentation", "Legacy FTP Decommission Plan", "SharePoint/Projects/FTP", "2026-01-05", "IT Manager", "Pending verification"],
        ["EVD-824.1-018", "4.1 VPN", "Configuration file", "OpenVPN Server Configuration", "Vault/configs/vpn", "2025-11-01", "Network Team", "Verified"],
        ["EVD-824.1-019", "4.1 VPN", "Documentation", "VPN MFA Enforcement Policy", "SharePoint/Policies/VPN", "2025-08-15", "Security Team", "Verified"],
        ["EVD-824.1-020", "4.1 VPN", "Configuration file", "WireGuard Configuration - Executive VPN", "Vault/configs/wireguard", "2025-12-01", "Security Team", "Verified"],
        ["EVD-824.1-021", "4.2 SSH", "Audit log", "SSH Bastion Host Audit Logs", "SIEM/ssh-bastion-logs", "2026-01-10", "Security Team", "Verified"],
        ["EVD-824.1-022", "4.2 SSH", "Documentation", "SSH Key Management Policy", "SharePoint/Policies/SSH", "2025-07-01", "Security Team", "Verified"],
        ["EVD-824.1-023", "4.2 SSH", "Certificate inventory", "Production Server SSH Key Inventory", "SharePoint/Inventory/SSH-Keys", "2026-01-08", "SysAdmin", "Verified"],
        ["EVD-824.1-024", "4.3 RDP", "Configuration file", "RDP NLA Configuration - Windows Servers", "SharePoint/GPO/RDP", "2025-10-15", "Windows Admin", "Verified"],
        ["EVD-824.1-025", "4.3 RDP", "Audit log", "RDP MFA Enforcement - Event Logs", "SIEM/rdp-mfa-events", "2026-01-09", "Security Team", "Verified"],
        ["EVD-824.1-026", "4.3 RDP", "Other", "Legacy RDP System Remediation Ticket", "Jira/SEC-1245", "2026-01-06", "Security Team", "Pending verification"],
        ["EVD-824.1-027", "5.1 API Security", "Configuration file", "API Gateway Security Configuration", "Git/configs/api-gateway", "2025-12-10", "API Team", "Verified"],
        ["EVD-824.1-028", "5.1 API Security", "Documentation", "OAuth 2.0 Implementation Documentation", "Confluence/API-Security", "2025-09-20", "Dev Team", "Verified"],
        ["EVD-824.1-029", "5.1 API Security", "Configuration file", "Service Mesh mTLS Configuration", "Git/k8s/istio-config", "2025-11-15", "Platform Team", "Verified"],
        ["EVD-824.1-030", "6.1 Database Connections", "Configuration file", "PostgreSQL TLS Configuration", "Vault/configs/postgres", "2025-10-20", "DBA Team", "Verified"],
        ["EVD-824.1-031", "6.1 Database Connections", "Documentation", "MongoDB X.509 Certificate Setup", "Confluence/MongoDB", "2025-11-05", "DBA Team", "Verified"],
        ["EVD-824.1-032", "6.1 Database Connections", "Documentation", "SQL Server Kerberos Authentication", "SharePoint/SQL-Security", "2025-09-30", "DBA Team", "Verified"],
        ["EVD-824.1-033", "6.2 Wireless Networks", "Compliance report", "Corporate WiFi WPA3 Deployment Report", "SharePoint/Projects/WiFi", "2025-08-01", "Network Team", "Verified"],
        ["EVD-824.1-034", "6.2 Wireless Networks", "Configuration file", "RADIUS Server Configuration", "Vault/configs/radius", "2025-10-10", "Network Team", "Verified"],
        ["EVD-824.1-035", "6.2 Wireless Networks", "Certificate inventory", "Wireless Access Point Inventory", "SharePoint/Inventory/WiFi", "2026-01-08", "Network Team", "Verified"],
        ["EVD-824.1-036", "7.1 Cloud Transmission", "Audit log", "AWS CloudTrail - VPC Flow Logs", "S3/cloudtrail/2026-01", "2026-01-10", "Cloud Team", "Verified"],
        ["EVD-824.1-037", "7.1 Cloud Transmission", "Compliance report", "Azure Monitor - Network Security Logs", "Azure/Monitor/Security", "2026-01-10", "Cloud Team", "Verified"],
        ["EVD-824.1-038", "7.1 Cloud Transmission", "Audit log", "GCP Cloud Audit Logs - Network", "GCP/Logging/Audit", "2026-01-10", "Cloud Team", "Verified"],
        ["EVD-824.1-039", "7.1 Cloud Transmission", "Documentation", "Multi-Cloud VPN Architecture Diagram", "SharePoint/Diagrams/Network", "2025-12-01", "Network Architect", "Verified"],
        ["EVD-824.1-040", "7.1 Cloud Transmission", "Screenshot", "CDN TLS Configuration - CloudFront", "SharePoint/Evidence/CDN", "2026-01-09", "Cloud Team", "Verified"],
    ]

    count = safely_write_data(ws, 5, evidence_data)  # Start at row 5 (after headers at row 4)
    logger.info("    Evidence Register: %d evidence documents", count)

    return count


def populate_approval_signoff(wb):
    """Populate Approval Sign-Off with workflow data."""

    if "Approval Sign-Off" not in wb.sheetnames:
        return 0

    ws = wb["Approval Sign-Off"]

    # Assessment completion
    today = datetime.now()
    ws["B5"] = "John Smith"  # Assessor Name
    ws["B6"] = "Senior Security Engineer"  # Role
    ws["B7"] = "Information Security"  # Department
    ws["B8"] = "john.smith@company.com"  # Email
    ws["B9"] = today.strftime("%Y-%m-%d")  # Date

    # Technical Review
    ws["B14"] = "Sarah Johnson"
    ws["B15"] = (today + timedelta(days=2)).strftime("%Y-%m-%d")
    ws["B16"] = "Reviewed and verified all TLS configurations and certificate validations. Minor gaps identified in legacy systems."

    # Security Review
    ws["B21"] = "Michael Chen"
    ws["B22"] = (today + timedelta(days=3)).strftime("%Y-%m-%d")
    ws["B23"] = "Security controls are adequate. Priority items: Legacy FTP migration and contractor VPN upgrade."

    # Management Approval
    ws["B28"] = "Jennifer Williams"
    ws["B29"] = (today + timedelta(days=5)).strftime("%Y-%m-%d")
    ws["B30"] = "Approved with conditions"
    ws["B31"] = "Approved pending completion of high-priority remediation items by Q2 2026."

    # Next Review
    ws["B36"] = (today + timedelta(days=90)).strftime("%Y-%m-%d")
    ws["B37"] = "Sarah Johnson"
    ws["B38"] = "Q2 2026 reassessment after remediation of critical gaps"

    logger.info("    Approval Sign-Off: Complete workflow populated")

    return 1


def main():
    """Main function to populate the Data Transmission assessment workbook."""
    try:
        if len(sys.argv) < 2:
            logger.error("Usage: python3 populate_a824_1_data_transmission.py <workbook.xlsx>")
            return 1

        filepath = sys.argv[1]

        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.8.24.1 - Data Transmission Assessment")
        logger.info("Comprehensive Data Population for CISO Presentation")
        logger.info("=" * 80)

        try:
            wb = load_workbook(filepath)
        except Exception as e:
            logger.error("Error loading workbook: %s", e)
            return 1

        logger.info("Loading: %s", filepath)
        logger.info("Sheets found: %d", len(wb.sheetnames))

        # Populate all sections
        logger.info("[1/3] Populating 12 Assessment Sheets...")
        assessment_count = populate_assessment_sheets(wb)

        logger.info("[2/3] Populating Evidence Register...")
        evidence_count = populate_evidence_register(wb)

        logger.info("[3/3] Populating Approval Sign-Off...")
        approval_count = populate_approval_signoff(wb)

        # Save workbook
        try:
            wb.save(filepath)
            logger.info("Saved: %s", filepath)
        except Exception as e:
            logger.error("Error saving: %s", e)
            return 1

        # Summary
        logger.info("=" * 80)
        logger.info("DATA POPULATION COMPLETE")
        logger.info("=" * 80)
        logger.info("Summary:")
        logger.info("  - Assessment Entries: %d", assessment_count)
        logger.info("  - Evidence Documents: %d", evidence_count)
        logger.info("  - Approval Workflow: Complete")
        logger.info("  Total Data Points: %d", assessment_count + evidence_count)

        logger.info("Compliance Distribution:")
        logger.info("  - ✅ Compliant: ~75%% (production-ready)")
        logger.info("  - Partial: ~17%% (minor gaps)")
        logger.info("  - ❌ Non-Compliant: ~8%% (critical items)")

        logger.info("Ready for CISO Presentation:")
        logger.info("  - All 12 transmission categories populated")
        logger.info("  - Realistic evidence documentation")
        logger.info("  - Complete approval workflow")
        logger.info("  - Professional data quality")

        logger.info("=" * 80)

        return 0

    except Exception as e:
        logger.error("Population failed with exception: %s", str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (standardized: license, imports, logging, exit codes)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: v1.0 - AGPL-3.0/Commercial dual-license
# =============================================================================
