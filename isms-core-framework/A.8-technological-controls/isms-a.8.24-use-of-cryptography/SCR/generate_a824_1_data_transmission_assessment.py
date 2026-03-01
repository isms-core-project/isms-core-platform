#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.24.1 - Data Transmission Cryptography Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography
Assessment Domain 1 of 4: Data Transmission Cryptographic Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific data transmission infrastructure, cryptographic
standards, and assessment requirements.

Key customization areas:
1. Protocol implementations and versions (match your actual infrastructure)
2. Cipher suite requirements and scoring criteria (adapt to your risk profile)
3. Certificate management systems (specific to your PKI infrastructure)
4. Integration points and data flows (based on your network architecture)
5. Compliance thresholds (aligned with your regulatory requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Use of Cryptography Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
cryptographic controls protecting data in transit across all network transmission
scenarios.

**Purpose:**
Enables systematic assessment of encryption implementation for data transmission
against ISO 27001:2022 Control A.8.24 requirements, supporting evidence-based
validation of cryptographic protection during network communication.

**Assessment Scope:**
- TLS/SSL protocol versions and cipher suites
- VPN encryption standards and implementations
- API communication security (REST, SOAP, GraphQL)
- File transfer protocols (SFTP, FTPS, SCP)
- Email encryption (S/MIME, PGP, TLS)
- Database connection encryption
- Internal network segmentation encryption
- Certificate lifecycle management
- Deprecated protocol identification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and cryptographic standards
2. TLS/SSL Assessment - Web services, APIs, and HTTPS implementations
3. VPN Assessment - Site-to-site and remote access VPN encryption
4. File Transfer Assessment - Secure file transfer protocol implementations
5. Email Security Assessment - Email encryption and signing mechanisms
6. Database Encryption - Database connection and replication encryption
7. Certificate Management - PKI infrastructure and certificate lifecycle
8. Protocol Inventory - Comprehensive protocol usage across infrastructure
9. Gap Analysis - Non-compliant protocols and remediation requirements
10. Evidence Register - Audit evidence tracking and documentation
11. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with cryptographic standard dropdown lists
- Conditional formatting for protocol/cipher compliance status
- Automated gap identification for deprecated algorithms
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with vulnerability scanning results

**Integration:**
This assessment feeds into the A.8.24.5 Compliance Dashboard, which
consolidates data from all four cryptographic assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a824_1_data_transmission_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a824_1_data_transmission_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a824_1_data_transmission_assessment.py --date 20250115

Output:
    File: ISMS_A_8_24_1_Data_Transmission_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize cryptographic standards to match your risk profile
    2. Inventory all data transmission systems and protocols
    3. Complete protocol assessments for each system/service
    4. Validate cipher suites against current best practices
    5. Review certificate management lifecycle
    6. Conduct gap analysis for deprecated/weak protocols
    7. Define remediation actions with timelines
    8. Collect and link audit evidence (configs, scan results)
    9. Obtain stakeholder approvals
    10. Feed results into A.8.24.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.24
Assessment Domain:    1 of 4 (Data Transmission Cryptographic Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               ISMS Core Contributors
Created:              2024
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              AGPL-3.0-or-later OR Commercial (dual-licensed)

Related Documents:
    - ISMS-POL-A.8.24: Use of Cryptography Policy (Governance)
    - ISMS-IMP-A.8.24.1: Data Transmission Cryptography Implementation Guide
    - ISMS-IMP-A.8.24.2: Data Storage Cryptography Assessment (Domain 2)
    - ISMS-IMP-A.8.24.3: Authentication Cryptography Assessment (Domain 3)
    - ISMS-IMP-A.8.24.4: Key Management Assessment (Domain 4)
    - ISMS-IMP-A.8.24.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.24.1 specification
    - Supports comprehensive data transmission cryptography evaluation
    - Integrated with A.8.24.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Cryptographic Standards:**
Cryptographic algorithms and protocols evolve rapidly. Review industry standards
(NIST, BSI, ECRYPT) quarterly and update assessment criteria accordingly.
Deprecated algorithms (MD5, SHA-1, 3DES, RC4) must be identified and remediated.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of protocol versions and cipher suites.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- System endpoints and network topology
- Certificate details and PKI infrastructure
- Vulnerability information and security gaps

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new cryptographic vulnerabilities
- Semi-annually: Update cipher suite recommendations
- Annually: Complete reassessment of all systems
- Ad-hoc: When infrastructure changes or new threats emerge

**Quality Assurance:**
Have cryptography SMEs and network security engineers validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Ensure cryptographic standards align with applicable regulatory requirements:
- Payment processing: PCI DSS v4.0.1 cryptographic requirements
- Healthcare: HIPAA encryption standards
- Finance: Regional banking encryption requirements
- Government: Jurisdiction-specific cryptographic mandates

Customize assessment criteria to include regulatory-specific requirements.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.24.1"
WORKBOOK_NAME = "Data Transmission Cryptography Assessment"
CONTROL_ID = "A.8.24"
CONTROL_NAME = "Use of Cryptography"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches markdown subsections
    sheets = [
        "Instructions & Legend",
        "1.1 External HTTPS-TLS",
        "1.2 Internal HTTPS-TLS",
        "2.1 Email Encryption",
        "2.2 Digital Signatures",
        "3.1 File Transfer Protocols",
        "4.1 VPN",
        "4.2 SSH",
        "4.3 RDP",
        "5.1 API Security",
        "6.1 Database Connections",
        "6.2 Wireless Networks",
        "7.1 Cloud Transmission",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles



_STYLES = setup_styles()
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 2: COLUMN DEFINITIONS PER SUBSECTION
# ============================================================================

def get_column_definitions(section_key):
    """
    Return column definitions (name: width) for each subsection.
    Matches the markdown specification exactly.
    """
    column_defs = {
        "1.1_external_https": {
            "Service Description": 35,
            "Current TLS Version": 18,
            "Certificate Source (CA)": 22,
            "Certificate Validity": 16,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "1.2_internal_https": {
            "Service Description": 35,
            "Data Classification": 18,
            "Current TLS Version": 18,
            "Certificate Type": 20,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "2.1_email_encryption": {
            "Email System": 25,
            "Encryption Solution": 22,
            "Encryption Method": 20,
            "User Adoption Rate": 16,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "2.2_digital_signatures": {
            "Use Case": 30,
            "Signature Method": 20,
            "Certificate Source": 22,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "3.1_file_transfer": {
            "Transfer Method/System": 30,
            "Protocol Used": 18,
            "Authentication Method": 22,
            "Data Classification": 18,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "4.1_vpn": {
            "VPN Solution": 25,
            "Protocol": 18,
            "Encryption Algorithm": 20,
            "MFA Enabled": 14,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "4.2_ssh": {
            "System/Service": 30,
            "SSH Version": 14,
            "Authentication Method": 22,
            "Key Algorithm": 18,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "4.3_rdp": {
            "System/Environment": 30,
            "RDP Access Method": 22,
            "TLS Encryption": 16,
            "NLA Enabled": 14,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "5.1_api": {
            "API Name/Service": 30,
            "Authentication Method": 22,
            "TLS Version": 14,
            "API Key Management": 22,
            "Token Expiry": 16,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "6.1_database": {
            "Database System": 25,
            "Connection Encryption": 22,
            "Certificate Validation": 20,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "6.2_wireless": {
            "Network SSID": 25,
            "Encryption Standard": 20,
            "Authentication Method": 22,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "7.1_cloud": {
            "Cloud Provider/Service": 30,
            "Connection Method": 22,
            "Encryption": 18,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
    }
    
    return column_defs.get(section_key, {})


def get_example_values(section_key):
    """
    Return realistic example row values for each assessment section.
    Column order must match get_column_definitions() exactly.
    """
    examples = {
        "1.1_external_https": [
            "Corporate website (www.example.com)", "TLS 1.3", "Let's Encrypt",
            "90 days", "✅ Compliant", "/evidence/a824/1.1-ssllabs-website-20260218.pdf",
            "None", "❌ No",
        ],
        "1.2_internal_https": [
            "HR portal (hr.internal.example.com)", "Confidential", "TLS 1.2",
            "Internal CA", "✅ Compliant", "/evidence/a824/1.2-internal-tls-config-20260218.txt",
            "None", "❌ No",
        ],
        "2.1_email_encryption": [
            "Microsoft Exchange Online", "S/MIME", "End-to-end encryption",
            "85%", "⚠️ Partial", "/evidence/a824/2.1-smime-policy-20260218.pdf",
            "User adoption below 90% target", "✅ Yes",
        ],
        "2.2_digital_signatures": [
            "Contract signing", "S/MIME", "Internal CA",
            "✅ Compliant", "/evidence/a824/2.2-digital-sig-cert-20260218.p12",
            "None", "❌ No",
        ],
        "3.1_file_transfer": [
            "SFTP Server (files.example.com)", "SFTP", "SSH key + MFA",
            "Confidential", "✅ Compliant", "/evidence/a824/3.1-sftp-config-20260218.txt",
            "None", "❌ No",
        ],
        "4.1_vpn": [
            "Cisco AnyConnect", "IKEv2/IPSec", "AES-256",
            "✅ Yes", "✅ Compliant", "/evidence/a824/4.1-vpn-config-20260218.pdf",
            "None", "❌ No",
        ],
        "4.2_ssh": [
            "Linux production servers", "SSH 2.0", "Public key (RSA)",
            "RSA-4096", "✅ Compliant", "/evidence/a824/4.2-ssh-key-audit-20260218.txt",
            "None", "❌ No",
        ],
        "4.3_rdp": [
            "Windows admin servers", "VPN + RDP Gateway", "TLS 1.2",
            "✅ Yes", "✅ Compliant", "/evidence/a824/4.3-rdp-config-20260218.png",
            "None", "❌ No",
        ],
        "5.1_api": [
            "Customer REST API", "OAuth 2.0 / JWT", "TLS 1.3",
            "Vault (HashiCorp)", "1 hour", "✅ Compliant",
            "/evidence/a824/5.1-api-tls-scan-20260218.pdf", "None", "❌ No",
        ],
        "6.1_database": [
            "PostgreSQL 15 (FINDB01)", "TLS 1.2 (sslmode=verify-full)",
            "Internal CA (verify-full)", "✅ Compliant",
            "/evidence/a824/6.1-db-tls-config-20260218.txt", "None", "❌ No",
        ],
        "6.2_wireless": [
            "CORP-WIFI", "WPA3-Enterprise", "802.1X / EAP-TLS",
            "✅ Compliant", "/evidence/a824/6.2-wifi-controller-20260218.png",
            "None", "❌ No",
        ],
        "7.1_cloud": [
            "AWS S3 (data backups)", "HTTPS/TLS 1.3", "AES-256 (in transit + at rest)",
            "✅ Compliant", "/evidence/a824/7.1-aws-config-20260218.pdf",
            "None", "❌ No",
        ],
    }
    return examples.get(section_key, [])
# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab for applicable systems/services.', '2. Use dropdown menus for standardised entries (Status, Remediation, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Check all applicable items in the Compliance Checklist for each section.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_assessment_sheet(ws, styles, section_title, policy_ref, question, 
                           section_key, checklist_items, additional_details):
    """
    Generic assessment sheet creator.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "1.1 HTTPS/TLS - External Web Services"
        policy_ref: policy requirement text
        question: assessment question
        section_key: key for column definitions (e.g., "1.1_external_https")
        checklist_items: list of compliance checklist items
        additional_details: list of (label, dropdown_options) tuples for Additional Details section
    """
    columns = get_column_definitions(section_key)
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{section_title.upper()}\nPolicy Requirement: {policy_ref}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.8.24: Use of Cryptography"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ---------- ASSESSMENT QUESTION ----------
    ws.merge_cells("A3:H3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ---------- RESPONSE DROPDOWN ----------
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"✅ Yes,❌ No,⚠️ Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    # ---------- COLUMN HEADERS ----------
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ---------- EXAMPLE ROW ----------
    example_row = 7
    example_values = get_example_values(section_key)
    # Pad to column count if needed (safety fallback)
    while len(example_values) < len(columns):
        example_values.append("")
    for col_idx, value in enumerate(example_values[:len(columns)], start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------- DATA ENTRY ROWS (8-30) ----------
    start_row = 8
    end_row = 30
    status_col_idx = list(columns.keys()).index("Status") + 1
    remediation_col_idx = list(columns.keys()).index("Remediation Needed") + 1
    
    for r in range(start_row, end_row + 1):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Status dropdown
    dv_status = DataValidation(type="list", 
                               formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"', 
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(start_row, end_row + 1):
        dv_status.add(ws.cell(row=r, column=status_col_idx))

    # Remediation Needed dropdown
    dv_rem = DataValidation(type="list", formula1='"✅ Yes,❌ No"', allow_blank=False)
    ws.add_data_validation(dv_rem)
    for r in range(start_row, end_row + 1):
        dv_rem.add(ws.cell(row=r, column=remediation_col_idx))

    ws.freeze_panes = "A7"

    next_row = end_row + 2

    # ---------- ADDITIONAL DETAILS ----------
    if additional_details:
        ws[f"A{next_row}"] = "Additional Details"
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1
        
        for label, options in additional_details:
            ws[f"A{next_row}"] = label
            ws[f"A{next_row}"].font = Font(bold=True)
            ws.merge_cells(f"B{next_row}:D{next_row}")
            ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{next_row}"].border = styles["border"]
            ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]
            
            if options:
                dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=False)
                ws.add_data_validation(dv)
                dv.add(ws[f"B{next_row}"])
            
            next_row += 1
        
        next_row += 1

    # ---------- COMPLIANCE CHECKLIST ----------
    if checklist_items:
        ws[f"A{next_row}"] = "COMPLIANCE CHECKLIST"
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1

        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = "Requirement"
        ws[f"C{next_row}"] = "Status"
        for col in ["A", "B", "C"]:
            ws[f"{col}{next_row}"].font = Font(bold=True)
        next_row += 1

        checklist_start = next_row
        for item in checklist_items:
            ws[f"A{next_row}"] = "☐"
            ws[f"B{next_row}"] = item
            ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
            ws[f"C{next_row}"].border = styles["border"]

            dv_chk = DataValidation(type="list", formula1='"✅ Yes,❌ No,N/A"', allow_blank=False)
            ws.add_data_validation(dv_chk)
            dv_chk.add(ws[f"C{next_row}"])
            next_row += 1

        # Checklist score
        ws[f"A{next_row}"] = "Checklist Score:"
        ws[f"A{next_row}"].font = Font(bold=True)
        ws[f"B{next_row}"] = (
            f'=IFERROR(ROUND(COUNTIF(C{checklist_start}:C{next_row-1},"✅ Yes")/'
            f'COUNTA(C{checklist_start}:C{next_row-1})*100,0)&"%","0%")'
        )
        ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
        next_row += 2

    # ---------- EXCEPTION/DEVIATION BLOCK ----------
    ws.merge_cells(f"A{next_row}:H{next_row}")
    ws[f"A{next_row}"] = "EXCEPTION / DEVIATION DOCUMENTATION (Complete if Status = Partial or Non-Compliant)"
    ws[f"A{next_row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{next_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{next_row}"].alignment = Alignment(horizontal="center", vertical="center")

    next_row += 1
    exception_fields = [
        ("Formal exception request submitted:", "✅ Yes,❌ No"),
        ("Exception ID:", ""),
        ("Risk acceptance documented:", "✅ Yes,❌ No"),
        ("Risk ID:", ""),
        ("Compensating Controls (summary):", ""),
        ("☐ Network segmentation / firewall restrictions", ""),
        ("☐ Enhanced monitoring / alerting", ""),
        ("☐ IP allowlisting / access restrictions", ""),
        ("☐ Other (describe):", ""),
        ("Remediation actions required:", ""),
        ("Responsible person:", ""),
        ("Target completion date:", ""),
        ("Budget required:", "✅ Yes,❌ No,⚠️ Unknown"),
    ]

    for label, options in exception_fields:
        ws[f"A{next_row}"] = label
        ws[f"A{next_row}"].font = Font(bold=True)
        ws.merge_cells(f"B{next_row}:D{next_row}")
        ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{next_row}"].border = styles["border"]
        ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]

        if options:
            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=False)
            ws.add_data_validation(dv)
            dv.add(ws[f"B{next_row}"])

        next_row += 1

    # ---------- NOTES ----------
    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row}")
    ws[f"A{next_row}"] = "ADDITIONAL NOTES / COMMENTS"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)

    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row+5}")
    ws[f"A{next_row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{next_row}"].border = styles["border"]
    ws[f"A{next_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # GS-AS-011: Borders on all merged cell ranges
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for _mr in list(ws.merged_cells.ranges):
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                ws.cell(row=_r, column=_c).border = _border

    return (start_row, end_row, status_col_idx)
# ============================================================================
# SECTION 5: INDIVIDUAL ASSESSMENT SHEET DEFINITIONS
# ============================================================================

def create_1_1_external_https_tls(ws, styles):
    """1.1 External HTTPS/TLS - matches markdown spec exactly."""
    checklist = [
        "TLS 1.3 preferred OR TLS 1.2 minimum",
        "Valid certificates from trusted public CA",
        "Certificate validity: ≤398d (until Mar 2026), ≤200d (Mar 2026+), ≤100d (Mar 2027+), ≤47d (Mar 2029+)",
        "Self-signed certificates NOT used in production",
        "HTTP automatically redirects to HTTPS",
        "HSTS header configured",
        "Strong cipher suites configured",
        "Weak protocols disabled (TLS 1.1, 1.0, SSL)",
        "Perfect Forward Secrecy (PFS) enabled",
        "Certificate expiration alerts (≥ 30 days)",
        "Automation infrastructure ready for 47-day certificate lifecycle (by Mar 2029)",
    ]
    
    additional_details = [
        ("Number of external web services:", ""),
        ("Certificate expiration monitoring configured:", "✅ Yes,❌ No"),
        ("Automated certificate renewal (REQUIRED for public CAs by Mar 2026):", "✅ Yes,❌ No,⏳ Planned"),
        ("Certificate inventory maintained:", "✅ Yes,❌ No"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1.1 HTTPS/TLS - External Web Services",
        policy_ref="All externally accessible web services MUST use TLS encryption with valid certificates from trusted CAs (Policy Section 2.2.1)",
        question="Does your organisation have external-facing web services or websites?",
        section_key="1.1_external_https",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_1_2_internal_https_tls(ws, styles):
    """1.2 Internal HTTPS/TLS - matches markdown spec."""
    checklist = [
        "TLS 1.2+ for services with Confidential/Restricted data",
        "Valid certificates (internal CA acceptable)",
        "Services with non-sensitive data: Risk assessed and documented",
        "Certificate inventory maintained",
        "Certificate expiration monitoring configured",
    ]
    
    additional_details = [
        ("Number of internal web services:", ""),
        ("Data classification(s) handled:", "Public,Internal,Confidential,Restricted"),
        ("Internal CA in use:", "✅ Yes,❌ No"),
        ("Certificate management process documented:", "✅ Yes,❌ No"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1.2 HTTPS/TLS - Internal Web Services",
        policy_ref="Internal web services containing sensitive data (Confidential or Restricted classification) MUST use TLS (Policy Section 2.2.1)",
        question="Does your organisation have internal web services (intranet, internal portals, internal APIs)?",
        section_key="1.2_internal_https",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_2_1_email_encryption(ws, styles):
    """2.1 Email Encryption - matches markdown spec."""
    checklist = [
        "S/MIME or PGP/GPG encryption available",
        "Users trained on when/how to encrypt sensitive emails",
        "PKI infrastructure in place for S/MIME",
        "Opportunistic TLS enabled for mail server connections",
        "STARTTLS enabled for SMTP",
    ]
    
    additional_details = [
        ("Encryption solution in use:", "S/MIME,PGP/GPG,TLS-only,Other"),
        ("PKI infrastructure for email:", "Implemented,Planned,Not implemented"),
        ("User training provided:", "✅ Yes,❌ No"),
        ("Encryption enforced via policy/DLP:", "✅ Yes,❌ No"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2.1 Email Encryption",
        policy_ref="Emails containing classified information (Confidential or Restricted) MUST be encrypted when sent externally (Policy Section 2.2.2)",
        question="Does your organisation send emails containing classified information to external parties?",
        section_key="2.1_email_encryption",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_2_2_digital_signatures(ws, styles):
    """2.2 Digital Signatures - matches markdown spec."""
    checklist = [
        "Digital signatures available for required use cases",
        "Email certificates issued from trusted CA or internal PKI",
        "Certificate validity ≤ 1 year",
        "Users trained on digital signature usage",
        "Non-repudiation implemented for legally binding transactions — digital signatures with qualified certificates where required",
        "Audit log integrity protected — cryptographic checksums (HMAC, digital signatures) applied to audit logs to detect tampering",
        "Time-stamping implemented — qualified electronic timestamps applied to critical evidence and transaction records",
    ]
    
    additional_details = [
        ("Digital signatures used for:", "All emails,Legal docs only,Executives only,Not used"),
        ("Certificate source:", "Public CA,Internal PKI,Both"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2.2 Digital Signatures",
        policy_ref="Digital signatures RECOMMENDED for all external emails, REQUIRED for legal/financial/official communications (Policy Section 2.2.2)",
        question="Does your organisation use digital signatures for email?",
        section_key="2.2_digital_signatures",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_3_1_file_transfer(ws, styles):
    """3.1 File Transfer Protocols - matches markdown spec."""
    checklist = [
        "SFTP, FTPS, or HTTPS used for sensitive file transfers",
        "Plain FTP NOT used for sensitive data",
        "Strong authentication configured (key-based preferred)",
        "MFA required for external file transfer",
        "File transfer logging enabled",
        "SSH keys rotated annually (if SFTP used)",
    ]
    
    additional_details = [
        ("Approved protocols in use:", "SFTP,FTPS,HTTPS,SCP,Other"),
        ("Prohibited protocols detected:", "FTP,TFTP,None"),
        ("Authentication:", "Password,Key-based,MFA,Certificate"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="3.1 Secure File Transfer",
        policy_ref="File transfers containing sensitive data MUST use encrypted protocols (SFTP, FTPS, HTTPS). Unencrypted FTP is PROHIBITED (Policy Section 2.2.3)",
        question="Does your organisation transfer files containing sensitive data to/from external parties or between systems?",
        section_key="3.1_file_transfer",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_4_1_vpn(ws, styles):
    """4.1 VPN - matches markdown spec."""
    checklist = [
        "Approved VPN protocol (IPsec/IKEv2, WireGuard, OpenVPN with TLS 1.2+)",
        "AES-256 or ChaCha20 encryption",
        "Perfect Forward Secrecy (PFS) enabled",
        "MFA required for all VPN connections",
        "Certificate-based authentication (preferred) OR strong pre-shared key",
        "VPN session timeout configured (≤30 minutes idle)",
        "Split-tunneling disabled (or documented exception)",
        "VPN access logs retained and reviewed",
    ]
    
    additional_details = [
        ("VPN protocol in use:", "IPsec/IKEv2,WireGuard,OpenVPN,Other"),
        ("Encryption algorithm:", ""),
        ("Number of VPN users:", ""),
        ("MFA solution:", "TOTP,Push notification,SMS,Hardware token,None"),
        ("Split-tunneling:", "Disabled,Enabled (with justification)"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4.1 VPN (Virtual Private Network)",
        policy_ref="All remote access to organisational networks MUST use encrypted VPN with approved protocols (IPsec, WireGuard, OpenVPN). MFA REQUIRED (Policy Section 2.2.4)",
        question="Does your organisation provide remote access via VPN?",
        section_key="4.1_vpn",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_4_2_ssh(ws, styles):
    """4.2 SSH - matches markdown spec."""
    checklist = [
        "SSH protocol version 2 only (SSHv1 disabled)",
        "Key-based authentication (password auth disabled preferred)",
        "Minimum key length: RSA 2048-bit or Ed25519",
        "Root login disabled",
        "SSH keys rotated annually",
        "Unused SSH keys removed",
        "Strong algorithms configured (per Policy Appendix A)",
        "Weak algorithms disabled (DSA, MD5, SHA-1)",
    ]
    
    additional_details = [
        ("SSH protocol version:", "SSHv2 only,SSHv1 (legacy - prohibited)"),
        ("Authentication method:", "Key-based only,Password allowed,Both"),
        ("SSH key types in use:", "Ed25519,RSA 3072+,RSA 2048,ECDSA,Other"),
        ("Root login via SSH:", "Disabled,Enabled"),
        ("SSH key rotation schedule:", "Annual,On personnel change,No rotation"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4.2 SSH (Secure Shell)",
        policy_ref="SSH REQUIRED for all administrative and remote terminal access. SSH protocol version 2 REQUIRED, password authentication SHOULD be disabled (Policy Section 2.2.4)",
        question="Does your organisation use SSH for remote system administration?",
        section_key="4.2_ssh",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_4_3_rdp(ws, styles):
    """4.3 RDP - matches markdown spec."""
    checklist = [
        "RDP accessed through VPN or jump host (NOT directly exposed)",
        "TLS encryption configured",
        "Network Level Authentication (NLA) enabled",
        "RDP encryption level set to 'High'",
        "MFA required for production system access",
        "RDP session recording (recommended for privileged access)",
    ]
    
    additional_details = [
        ("RDP access method:", "VPN required,Jump host/bastion,Direct (prohibited),Zero-trust gateway"),
        ("Network Level Authentication (NLA):", "Enabled,Disabled"),
        ("MFA for RDP:", "Required,Optional,Not implemented"),
        ("RDP encryption level:", "High,Medium,Low"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4.3 RDP (Remote Desktop Protocol)",
        policy_ref="RDP connections MUST be encrypted using TLS. RDP MUST NOT be directly exposed to the Internet (Policy Section 2.2.4)",
        question="Does your organisation use RDP for remote access?",
        section_key="4.3_rdp",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_5_1_api_security(ws, styles):
    """5.1 API Security - matches markdown spec."""
    checklist = [
        "All API endpoints use HTTPS (TLS 1.2+)",
        "HTTP endpoints disabled or blocked",
        "OAuth 2.0 with PKCE (for user-facing APIs)",
        "API keys have ≥256-bit entropy",
        "API keys rotate every 90 days",
        "API keys stored in secrets manager (NOT in code/config)",
        "Access tokens expire within 1 hour",
        "Refresh tokens expire within 24 hours",
        "Rate limiting implemented per API key",
        "API keys NOT passed in URL query parameters",
    ]
    
    additional_details = [
        ("Number of APIs:", ""),
        ("API authentication:", "OAuth 2.0,API keys,Mutual TLS,JWT,Basic auth,None"),
        ("API key entropy:", "≥256 bits,<256 bits,Unknown"),
        ("API key rotation:", "Automated (90 days),Manual,No rotation"),
        ("API keys stored in:", "Secrets manager,Environment variables,Code (prohibited),Config files"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="5.1 API Security",
        policy_ref="API endpoints MUST use HTTPS with TLS 1.2+. API authentication MUST use OAuth 2.0, API keys (256-bit entropy), or mutual TLS (Policy Section 2.2.5)",
        question="Does your organisation expose APIs (REST, SOAP, GraphQL, etc.)?",
        section_key="5.1_api",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_6_1_database_connections(ws, styles):
    """6.1 Database Connections - matches markdown spec."""
    checklist = [
        "Database connections encrypted (TLS/SSL)",
        "Certificate validation enabled (not 'trust any certificate')",
        "Self-signed certificates only for internal databases (with proper CA)",
        "Unencrypted connections disabled or documented exception",
    ]
    
    additional_details = [
        ("Database systems in use:", "PostgreSQL,MySQL,MSSQL,Oracle,MongoDB,Other"),
        ("Connection encryption enforced:", "✅ Yes,❌ No,⚠️ Mixed"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="6.1 Database Connections",
        policy_ref="Database connections MUST use encrypted protocols (TLS for PostgreSQL/MySQL, encrypted connections for MSSQL) (Policy Section 2.2.6)",
        question="Does your organisation have applications connecting to databases?",
        section_key="6.1_database",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_6_2_wireless_networks(ws, styles):
    """6.2 Wireless Networks - matches markdown spec."""
    checklist = [
        "WPA3-Enterprise or WPA2-Enterprise for corporate networks",
        "802.1X with EAP-TLS (certificate-based) preferred",
        "WPA2-Personal only with strong passphrase (≥20 characters)",
        "WEP and WPA (original) NOT used",
        "Guest wireless isolated from corporate network",
        "WiFi passwords rotated quarterly (for PSK networks)",
    ]
    
    additional_details = [
        ("Corporate WiFi encryption:", "WPA3-Enterprise,WPA2-Enterprise,WPA2-Personal,Other"),
        ("Guest WiFi:", "Isolated network,Captive portal,Open (no encryption)"),
        ("802.1X authentication:", "Implemented,Planned,Not implemented"),
        ("WiFi password strength:", "≥20 characters,<20 characters"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="6.2 Wireless Networks",
        policy_ref="Wireless networks MUST use WPA3-Enterprise or WPA2-Enterprise minimum. WEP and WPA (original) PROHIBITED (Policy Section 2.2.6)",
        question="Does your organisation have wireless networks?",
        section_key="6.2_wireless",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_7_1_cloud_transmission(ws, styles):
    """7.1 Cloud Data Transmission - matches markdown spec."""
    checklist = [
        "TLS 1.2+ for all cloud API connections",
        "Private connectivity for Confidential/Restricted data (preferred)",
        "Cloud provider native encryption enabled",
        "Data encrypted in transit to/from cloud",
    ]
    
    additional_details = [
        ("Connection type:", "Public internet (TLS),VPN,Private Link/PrivateLink,Direct Connect/ExpressRoute"),
        ("Data classification transmitted:", "Public,Internal,Confidential,Restricted"),
        ("Cloud providers used:", "AWS,Azure,GCP,Office 365,Salesforce,Other"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="7.1 Cloud Data Transmission",
        policy_ref="Connections to cloud provider APIs MUST use TLS 1.2+. Private connectivity options (PrivateLink, Private Link, Private Service Connect) RECOMMENDED for high-volume or sensitive data (Policy Section 2.2.7)",
        question="Does your organisation use cloud services (AWS, Azure, GCP, SaaS)?",
        section_key="7.1_cloud",
        checklist_items=checklist,
        additional_details=additional_details,
    )
# ============================================================================
# SECTION 6: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard — TABLE 1/2/3, all tables uniform 7-col A:G width."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Row 1: Title ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font      = styles["header"]["font"]
    ws["A1"].fill      = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border

    # --- Row 2: ISO reference ---
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.8.24 | Data Transmission Cryptography"
    ws["A2"].font      = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    # --- TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW ---
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    c.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="003366")
    #c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    for col_idx, h in enumerate(["Assessment Area", "Total Items", "Compliant", "Partial",
                                  "Non-Compliant", "N/A", "Compliance %"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill      = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    areas = [
        ("1.1 External HTTPS/TLS",    "1.1 External HTTPS-TLS",      5),
        ("1.2 Internal HTTPS/TLS",    "1.2 Internal HTTPS-TLS",      5),
        ("2.1 Email Encryption",      "2.1 Email Encryption",        5),
        ("2.2 Digital Signatures",    "2.2 Digital Signatures",      4),
        ("3.1 File Transfer",         "3.1 File Transfer Protocols", 5),
        ("4.1 VPN",                   "4.1 VPN",                     5),
        ("4.2 SSH",                   "4.2 SSH",                     5),
        ("4.3 RDP",                   "4.3 RDP",                     5),
        ("5.1 API Security",          "5.1 API Security",            6),
        ("6.1 Database Connections",  "6.1 Database Connections",    4),
        ("6.2 Wireless Networks",     "6.2 Wireless Networks",       4),
        ("7.1 Cloud Transmission",    "7.1 Cloud Transmission",      4),
    ]
    row += 1
    start_data_row = row
    for label, sheet, status_col in areas:
        ws.cell(row=row, column=1, value=label)
        sc = get_column_letter(status_col)
        rng = f"'{sheet}'!{sc}8:{sc}69"
        ws.cell(row=row, column=2, value=f"=COUNTA({rng})")
        ws.cell(row=row, column=3, value=f'=COUNTIF({rng},"*Compliant")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({rng},"*Partial")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({rng},"*Non-Compliant")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({rng},"N/A")')
        ws.cell(row=row, column=7,
                value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font  = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill                 = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border               = border
    for col in range(2, 7):
        c       = ws.cell(row=total_row, column=col)
        c.value = f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{total_row - 1})"
        c.font  = Font(name="Calibri", size=10, bold=True)
        c.fill  = PatternFill("solid", fgColor="D9D9D9")
        c.border = border
    c       = ws.cell(row=total_row, column=7)
    c.value = (f'=IF((B{total_row}-F{total_row})=0,"0%",'
               f'ROUND(C{total_row}/(B{total_row}-F{total_row})*100,1)&"%")')
    c.font   = Font(name="Calibri", size=12, bold=True, color="000000")
    c.fill   = PatternFill("solid", fgColor="D9D9D9")
    c.border = border

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 15

    _kpis = [
        ("TLS 1.2+ Coverage – External", "100%"),
        ("TLS 1.2+ Coverage – Internal", "≥95%"),
        ("Email Encryption – Sensitive Data", "100%"),
        ("VPN Compliance Rate", "100%"),
        ("SSH Key Rotation Compliance", "≥90%"),
        ("API mTLS Coverage", "≥95%"),
        ("Database Encryption in Transit", "100%"),
        ("Cloud Transmission Encryption", "100%"),
        ("Certificates Expiring < 30 Days (count)", "0"),
        ("Evidence Documentation Rate", "100%"),
    ]
    # --- TABLE 2: KEY METRICS ---
    row = total_row + 2
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    #c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    for col_idx, h in enumerate(["KPI", "Current Value", "Target", "Status",
                                  "Last Updated", "Owner", "Notes"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    dv_kpi_sts = DataValidation(type="list",
        formula1='"\u2705 On Target,\u26a0\ufe0f At Risk,\u274c Below Target"',
        allow_blank=True)
    ws.add_data_validation(dv_kpi_sts)

    row += 1
    for kpi, target in _kpis:
        ws.cell(row=row, column=1, value=kpi).border = border
        c2 = ws.cell(row=row, column=2)
        c2.border = border
        ws.cell(row=row, column=3, value=target).border = border
        c4 = ws.cell(row=row, column=4)
        c4.border = border
        dv_kpi_sts.add(c4)
        for col in range(5, 8):
            ws.cell(row=row, column=col).border = border
        row += 1
    r = row

    # 2 bordered buffer rows + blank gap before TABLE 3 (Gold Standard)
    for _buf in range(2):
        for _col in range(1, 8):
            ws.cell(row=r, column=_col).border = border
        r += 1

    # --- TABLE 3: DOMAIN COMPLIANCE SUMMARY ---
    row = r + 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: DOMAIN COMPLIANCE SUMMARY")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    t3_headers = [
        "Transmission Domain",
        "Non-Compliant",
        "Partial",
        "Compliant",
        "N/A",
        "Total Items",
        "Compliance %",
    ]
    for col_idx, h in enumerate(t3_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill      = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    # Domain rows: (label, sheet_name, status_col_letter)
    t3_domains = [
        ("1.1 External HTTPS/TLS",   "1.1 External HTTPS-TLS",      "E"),
        ("1.2 Internal HTTPS/TLS",   "1.2 Internal HTTPS-TLS",      "E"),
        ("2.1 Email Encryption",     "2.1 Email Encryption",        "E"),
        ("2.2 Digital Signatures",   "2.2 Digital Signatures",      "D"),
        ("3.1 File Transfer",        "3.1 File Transfer Protocols", "E"),
        ("4.1 VPN",                  "4.1 VPN",                     "E"),
        ("4.2 SSH",                  "4.2 SSH",                     "E"),
        ("4.3 RDP",                  "4.3 RDP",                     "E"),
        ("5.1 API Security",         "5.1 API Security",            "F"),
        ("6.1 Database Connections", "6.1 Database Connections",    "D"),
        ("6.2 Wireless Networks",    "6.2 Wireless Networks",       "D"),
        ("7.1 Cloud Transmission",   "7.1 Cloud Transmission",      "D"),
    ]
    row += 1
    t3_start = row
    for label, sheet, sc in t3_domains:
        rng = f"\'{sheet}\'!{sc}8:{sc}69"
        non_c  = f'=COUNTIF({rng},"*Non-Compliant")'
        part   = f'=COUNTIF({rng},"*Partial")'
        compl  = f'=COUNTIF({rng},"*Compliant")'
        na_    = f'=COUNTIF({rng},"N/A")'
        total  = f'=COUNTA({rng})'
        pct_r  = row
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font   = Font(name="Calibri", size=10, color="000000")
        c1.fill   = PatternFill("solid", fgColor="FFFFCC")
        c1.border = border
        for col_idx, fml in enumerate([non_c, part, compl, na_, total], 2):
            cx = ws.cell(row=row, column=col_idx, value=fml)
            cx.font   = Font(name="Calibri", size=10, color="000000")
            cx.fill   = PatternFill("solid", fgColor="FFFFCC")
            cx.border = border
        c7 = ws.cell(row=row, column=7,
                     value=f'=IF((F{pct_r}-E{pct_r})=0,0,D{pct_r}/(F{pct_r}-E{pct_r}))')
        c7.font        = Font(name="Calibri", size=10, color="000000")
        c7.fill        = PatternFill("solid", fgColor="FFFFCC")
        c7.number_format = "0.0%"
        c7.border      = border
        row += 1

    # TOTAL row
    t3_end = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font  = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=1).fill   = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=row, column=1).border = border
    for col in range(2, 7):
        cx = ws.cell(row=row, column=col)
        cx.value  = f"=SUM({get_column_letter(col)}{t3_start}:{get_column_letter(col)}{t3_end})"
        cx.font   = Font(name="Calibri", size=10, bold=True)
        cx.fill   = PatternFill("solid", fgColor="D9D9D9")
        cx.border = border
    tot_row = row
    cx = ws.cell(row=row, column=7,
                 value=f'=IF((F{tot_row}-E{tot_row})=0,0,D{tot_row}/(F{tot_row}-E{tot_row}))')
    cx.font          = Font(name="Calibri", size=10, bold=True, color="000000")
    cx.fill          = PatternFill("solid", fgColor="D9D9D9")
    cx.number_format = "0.0%"
    cx.border        = border
    row += 1

    # Borders on all merged cell ranges (GS-AS-011)
    _t = Side(style="thin")
    _b = Border(left=_t, right=_t, top=_t, bottom=_t)
    for _mr in list(ws.merged_cells.ranges):
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                ws.cell(row=_r, column=_c).border = _b

    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create evidence register for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validation dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    # Sample row (row 5) — F2F2F2 grey with realistic example data
    _er_sample = [
        "EV-001", "1.1 External HTTPS-TLS", "Configuration file",
        "TLS/encryption configuration export for audit evidence",
        "/evidence/a824/config-export-2025.txt",
        "01.01.2025", "Security Team", "✅ Verified",
    ]
    for c, val in enumerate(_er_sample, start=1):
        cell = ws.cell(row=5, column=c, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if c == 1:
            cell.font = Font(italic=True, color="555555")
    dv_type.add(ws.cell(row=5, column=3))
    dv_ver.add(ws.cell(row=5, column=8))

    # Empty data rows (rows 6-105) — 100 FFFFCC input rows, NO pre-filled IDs
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(ws):
    """Create approval and sign-off sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.24.1 - Data Transmission Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "='Summary Dashboard'!G18"),
        ("Assessment Status:", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    for merge_range in list(ws.merged_cells.ranges):
        for _row in range(merge_range.min_row, merge_range.max_row + 1):
            for _col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=_row, column=_col).border = border
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 9: MAIN EXECUTION
# ============================================================================

def main() -> int:
    """
    Main execution function - orchestrates workbook creation.

    Returns:
        int: 0 on success, 1 on failure
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.24.1 - Data Transmission Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography")
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/16] Creating Instructions & Legend sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/16] Creating 1.1 External HTTPS/TLS sheet...")
        create_1_1_external_https_tls(wb["1.1 External HTTPS-TLS"], styles)

        logger.info("[3/16] Creating 1.2 Internal HTTPS/TLS sheet...")
        create_1_2_internal_https_tls(wb["1.2 Internal HTTPS-TLS"], styles)

        logger.info("[4/16] Creating 2.1 Email Encryption sheet...")
        create_2_1_email_encryption(wb["2.1 Email Encryption"], styles)

        logger.info("[5/16] Creating 2.2 Digital Signatures sheet...")
        create_2_2_digital_signatures(wb["2.2 Digital Signatures"], styles)

        logger.info("[6/16] Creating 3.1 File Transfer sheet...")
        create_3_1_file_transfer(wb["3.1 File Transfer Protocols"], styles)

        logger.info("[7/16] Creating 4.1 VPN sheet...")
        create_4_1_vpn(wb["4.1 VPN"], styles)

        logger.info("[8/16] Creating 4.2 SSH sheet...")
        create_4_2_ssh(wb["4.2 SSH"], styles)

        logger.info("[9/16] Creating 4.3 RDP sheet...")
        create_4_3_rdp(wb["4.3 RDP"], styles)

        logger.info("[10/16] Creating 5.1 API Security sheet...")
        create_5_1_api_security(wb["5.1 API Security"], styles)

        logger.info("[11/16] Creating 6.1 Database Connections sheet...")
        create_6_1_database_connections(wb["6.1 Database Connections"], styles)

        logger.info("[12/16] Creating 6.2 Wireless Networks sheet...")
        create_6_2_wireless_networks(wb["6.2 Wireless Networks"], styles)

        logger.info("[13/16] Creating 7.1 Cloud Transmission sheet...")
        create_7_1_cloud_transmission(wb["7.1 Cloud Transmission"], styles)

        logger.info("[14/16] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

        logger.info("[15/16] Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[16/16] Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        output_path = _wkbk_dir / OUTPUT_FILENAME
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"SUCCESS: {output_path}")
        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions & Legend")
        logger.info("  2) Fill yellow cells in each assessment sheet (1.1 - 7.1)")
        logger.info("  3) Check compliance checklists per section")
        logger.info("  4) Document exceptions/deviations as needed")
        logger.info("  5) Maintain Evidence Register entries")
        logger.info("  6) Review Summary Dashboard for compliance gaps")
        logger.info("  7) Complete Approval Sign-Off")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error(f"Permission denied: {e}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
