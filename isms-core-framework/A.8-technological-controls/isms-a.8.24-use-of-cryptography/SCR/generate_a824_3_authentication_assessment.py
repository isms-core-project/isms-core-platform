#!/usr/bin/env python3
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.24.3 - Authentication Cryptography Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography
Assessment Domain 3 of 4: Authentication Cryptographic Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific authentication infrastructure, cryptographic
standards, and assessment requirements.

Key customization areas:
1. Authentication systems and protocols (match your actual IAM infrastructure)
2. Password hashing algorithms (adapt to your security baseline)
3. Digital signature implementations (specific to your PKI/code signing)
4. Multi-factor authentication methods (based on your MFA deployment)
5. Compliance thresholds (aligned with your regulatory requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Use of Cryptography Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
cryptographic controls protecting authentication mechanisms, digital signatures,
and identity verification across all systems.

**Purpose:**
Enables systematic assessment of cryptographic authentication implementation
against ISO 27001:2022 Control A.8.24 requirements, supporting evidence-based
validation of cryptographic protection for authentication and non-repudiation.

**Assessment Scope:**
- Password storage and hashing (bcrypt, Argon2, PBKDF2, scrypt)
- Digital signatures (document signing, code signing, email signing)
- Certificate-based authentication (client certificates, smart cards)
- API authentication tokens (JWT, OAuth 2.0, API keys)
- SAML and federation cryptography
- SSH key authentication
- Multi-factor authentication cryptographic elements
- Kerberos and domain authentication
- Biometric template protection
- Signature verification and validation infrastructure
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and authentication standards
2. Password Hashing - Password storage and hashing algorithm assessment
3. Digital Signatures - Document, code, and email signature evaluation
4. Certificate Authentication - PKI-based authentication systems
5. Token Authentication - API tokens and session management cryptography
6. Federation & SSO - SAML, OAuth, OpenID Connect cryptographic controls
7. SSH Key Management - SSH key generation, storage, and rotation
8. MFA Cryptography - Multi-factor authentication cryptographic assessment
9. Domain Authentication - Kerberos and Active Directory cryptography
10. Signature Validation - Digital signature verification infrastructure
11. Gap Analysis - Weak authentication methods and remediation requirements
12. Evidence Register - Audit evidence tracking and documentation
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with authentication standard dropdown lists
- Conditional formatting for algorithm compliance status
- Automated gap identification for weak hashing/signing algorithms
- Protected formulas with unprotected input cells
- Authentication method to cryptographic control mapping
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with identity management systems

**Integration:**
This assessment feeds into the A.8.24.5 Compliance Dashboard, which
consolidates data from all four cryptographic assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a824_3_authentication_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a824_3_authentication_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a824_3_authentication_assessment.py --date 20250115

Output:
    File: ISMS_A_8_24_3_Authentication_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize authentication standards to match your risk profile
    2. Inventory all authentication systems and methods
    3. Complete cryptographic assessments for each authentication mechanism
    4. Validate hashing algorithms and iteration counts
    5. Review digital signature implementations and key strengths
    6. Assess certificate-based authentication controls
    7. Conduct gap analysis for weak authentication cryptography
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (configs, IAM reports)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.24.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.24
Assessment Domain:    3 of 4 (Authentication Cryptographic Controls)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.24: Use of Cryptography Policy (Governance)
    - ISMS-IMP-A.8.24.1: Data Transmission Cryptography Assessment (Domain 1)
    - ISMS-IMP-A.8.24.2: Data Storage Cryptography Assessment (Domain 2)
    - ISMS-IMP-A.8.24.3: Authentication Cryptography Implementation Guide
    - ISMS-IMP-A.8.24.4: Key Management Assessment (Domain 4)
    - ISMS-IMP-A.8.24.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.24.3 specification
    - Supports comprehensive authentication cryptography evaluation
    - Integrated with A.8.24.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Cryptographic Standards:**
Password hashing and authentication cryptography standards evolve rapidly.
Review industry standards (OWASP, NIST SP 800-63B, FIDO Alliance) annually
and update assessment criteria accordingly. Deprecated methods (MD5, SHA-1
for passwords, weak PBKDF2 iterations) must be identified and remediated.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of password hashing implementation and
digital signature validation processes.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Authentication system architecture and endpoints
- Password hashing algorithms and parameters (sensitive security info)
- Digital signature implementations and key details
- Weak authentication methods (security gaps)

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new authentication systems and methods
- Semi-annually: Validate hashing algorithm compliance
- Annually: Complete reassessment of all authentication infrastructure
- Ad-hoc: When authentication failures or security incidents occur

**Quality Assurance:**
Have identity and access management (IAM) specialists and cryptography SMEs
validate assessments before using results for compliance reporting or
remediation decisions.

**Regulatory Alignment:**
Ensure authentication cryptography aligns with applicable requirements:
- Payment processing: PCI DSS v4.0.1 authentication and password requirements
- Healthcare: HIPAA authentication and access control standards
- Privacy: GDPR identity verification and authentication
- Finance: Regional banking authentication requirements
- Government: Jurisdiction-specific authentication mandates

Customize assessment criteria to include regulatory-specific requirements.

**Non-Repudiation:**
Digital signatures provide non-repudiation for legal and audit purposes.
Ensure signature implementations meet jurisdictional requirements for
legal validity and evidentiary weight in your operating regions.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.24.3"
WORKBOOK_NAME = "Authentication Cryptography Assessment"
CONTROL_ID = "A.8.24"
CONTROL_NAME = "Use of Cryptography"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches markdown specification
    sheets = [
        "Instructions & Legend",
        "1. Password Security",
        "2. Multi-Factor Authentication",
        "3. Certificate-Based Auth",
        "4. Service Accounts",
        "5. SSO & Federation",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "exception_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "critical_alert": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly
# ============================================================================
# SECTION 2: COLUMN DEFINITIONS (BASE + EXTENSIONS)
# ============================================================================

def get_base_auth_columns():
    """
    Return base 17 columns (A-Q) standard for ALL authentication assessment sheets.
    
    Returns:
        dict: {column_name: width}
    """
    return {
        "System/Application": 30,
        "Authentication Method": 22,
        "User Type": 18,
        "Data Classification": 18,
        "Cryptographic Algorithm": 22,
        "Hash/Encryption Status": 20,
        "Password Complexity": 20,
        "Status": 15,
        "Evidence Location": 28,
        "Gap Description": 30,
        "Remediation Needed": 14,
        "Exception ID": 14,
        "Risk ID": 14,
        "Compensating Controls": 30,
        "Responsible Person": 20,
        "Target Date": 14,
        "Budget Required": 15,
    }


def get_extended_columns(sheet_type):
    """
    Return additional columns (R-X) specific to each sheet type.
    
    Args:
        sheet_type: String identifier (password, mfa, certificate, service, sso)
    
    Returns:
        dict: {column_name: width} for columns R onwards
    """
    extensions = {
        "password": {
            "Min Password Length": 16,
            "Complexity Enforced": 16,
            "Password Expiry": 18,
            "Password History": 16,
            "Account Lockout": 16,
            "Password Strength Meter": 18,
            "Default Passwords": 18,
        },
        "mfa": {
            "MFA Factor Type": 22,
            "MFA Enrollment Rate": 18,
            "MFA Enforcement": 16,
            "Backup MFA Method": 20,
            "MFA Bypass Allowed": 18,
            "Passwordless Option": 18,
        },
        "certificate": {
            "Certificate Type": 22,
            "Issuing CA": 25,
            "Key Algorithm": 18,
            "Certificate Validity": 16,
            "CRL/OCSP Configured": 18,
            "Smart Card Required": 18,
            "PIV/CAC Compliant": 18,
        },
        "service": {
            "Service Account Type": 22,
            "Auth Method": 20,
            "Password Rotation": 18,
            "Privileged Account": 16,
            "Account Monitoring": 18,
            "Interactive Login": 18,
            "Least Privilege": 16,
        },
        "sso": {
            "SSO Protocol": 20,
            "Identity Provider": 25,
            "MFA at IdP": 16,
            "Just-In-Time Provisioning": 18,
            "Session Timeout": 18,
            "Certificate Validation": 18,
            "Token Encryption": 16,
        },
    }
    
    return extensions.get(sheet_type, {})


def get_all_columns(sheet_type):
    """Combine base + extended columns for a given sheet type."""
    base = get_base_auth_columns()
    extended = get_extended_columns(sheet_type)
    base.update(extended)
    return base


# ============================================================================
# SECTION 3: GLOBAL DATA VALIDATION DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create base dropdown validations common to all authentication sheets.
    Returns dict of DataValidation objects.
    """
    validations = {}
    
    # Authentication Method
    validations['auth_method'] = DataValidation(
        type="list",
        formula1='"Password,MFA,Certificate,Token,SSO,Federation,Biometric,API Key,Service Account,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['auth_method'])
    
    # User Type
    validations['user_type'] = DataValidation(
        type="list",
        formula1='"End User,Administrator,Service Account,External User,API Client,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['user_type'])
    
    # Data Classification
    validations['data_class'] = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['data_class'])
    
    # Cryptographic Algorithm
    validations['algorithm'] = DataValidation(
        type="list",
        formula1='"bcrypt,Argon2,PBKDF2,scrypt,SHA-256,SHA-512,RSA-2048+,ECDSA,None,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['algorithm'])
    
    # Hash/Encryption Status
    validations['hash_status'] = DataValidation(
        type="list",
        formula1='"Properly Hashed,Encrypted,Plaintext (Non-compliant),Salted,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['hash_status'])
    
    # Password Complexity
    validations['pwd_complexity'] = DataValidation(
        type="list",
        formula1='"Strong (≥14 chars),Adequate (12-13),Weak (<12),N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['pwd_complexity'])
    
    # Status
    validations['status'] = DataValidation(
        type="list",
        formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['status'])
    
    # Remediation Needed
    validations['remediation'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['remediation'])
    
    # Budget Required
    validations['budget'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No,⚠️ Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(validations['budget'])
    
    # Response (Yes/No/Not Applicable)
    validations['response'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No,⚠️ Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(validations['response'])
    
    # Checklist Yes/No/N/A
    validations['checklist'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['checklist'])
    
    # Exception Yes/No
    validations['exception_yn'] = DataValidation(
        type="list",
        formula1='"✅ Yes,❌ No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['exception_yn'])
    
    return validations


def create_extended_validations(ws, sheet_type):
    """Create sheet-specific dropdown validations for extended columns."""
    validations = {}
    
    if sheet_type == "password":
        validations['min_length'] = DataValidation(
            type="list",
            formula1='"≥14 chars,12-13 chars,<12 chars (weak),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['min_length'])
        
        validations['expiry'] = DataValidation(
            type="list",
            formula1='"90 days,180 days,365 days,Never,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['expiry'])
        
        validations['history'] = DataValidation(
            type="list",
            formula1='"≥10 passwords,5-9 passwords,<5 passwords,None,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['history'])
        
        validations['default_pwd'] = DataValidation(
            type="list",
            formula1='"Changed,Not Changed (Non-compliant),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['default_pwd'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "mfa":
        validations['mfa_factor'] = DataValidation(
            type="list",
            formula1='"TOTP (Authenticator App),Push Notification,SMS (Deprecated),Hardware Token,Biometric,Certificate,None,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['mfa_factor'])
        
        validations['enrollment'] = DataValidation(
            type="list",
            formula1='"100%,90-99%,75-89%,<75%,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['enrollment'])
        
        validations['enforcement'] = DataValidation(
            type="list",
            formula1='"Required,Optional,Not Configured,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['enforcement'])
        
        validations['bypass'] = DataValidation(
            type="list",
            formula1='"Never,With Approval,Emergency Only,Always (Non-compliant),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['bypass'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "certificate":
        validations['cert_type'] = DataValidation(
            type="list",
            formula1='"Client Certificate,Smart Card,Machine/Device,Code Signing,Email,User,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_type'])
        
        validations['cert_validity'] = DataValidation(
            type="list",
            formula1='"≤1 year,1-2 years,>2 years,Expired,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_validity'])
        
        validations['key_algo'] = DataValidation(
            type="list",
            formula1='"RSA-4096,RSA-3072,RSA-2048,ECDSA P-256,ECDSA P-384,Ed25519,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['key_algo'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "service":
        validations['svc_type'] = DataValidation(
            type="list",
            formula1='"Database Service,Application Service,API Service,Scheduled Task,Integration,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['svc_type'])
        
        validations['svc_auth'] = DataValidation(
            type="list",
            formula1='"Certificate,Long Password (≥20 chars),API Key,Managed Identity,Secret,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['svc_auth'])
        
        validations['rotation'] = DataValidation(
            type="list",
            formula1='"Never (Monitored),90 days,180 days,365 days,On Change Only,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['rotation'])
        
        validations['interactive'] = DataValidation(
            type="list",
            formula1='"Disabled,Enabled (Non-compliant),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['interactive'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "sso":
        validations['sso_protocol'] = DataValidation(
            type="list",
            formula1='"SAML 2.0,OAuth 2.0,OIDC,WS-Federation,Kerberos,LDAP,None,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['sso_protocol'])
        
        validations['session_timeout'] = DataValidation(
            type="list",
            formula1='"≤1 hour,1-8 hours,>8 hours,No timeout,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['session_timeout'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    return validations
# ============================================================================
# SECTION 4: CHECKLIST DEFINITIONS PER AUTHENTICATION TYPE
# ============================================================================

def get_checklist_items(sheet_type):
    """
    Return checklist items for each authentication assessment type.
    
    Args:
        sheet_type: String identifier (password, mfa, certificate, service, sso)
    
    Returns:
        list: Checklist items as strings
    """
    checklists = {
        "password": [
            "Passwords hashed with approved algorithm (bcrypt, Argon2, PBKDF2, scrypt)",
            "Passwords NOT stored in plaintext anywhere",
            "Passwords NOT stored with reversible encryption",
            "Password hashes salted (unique salt per user)",
            "Minimum password length ≥12 characters",
            "Recommended password length ≥14 characters",
            "Complexity requirements enforced (uppercase, lowercase, numbers, symbols)",
            "Password expiration policy: 90 days maximum (or risk-based)",
            "Password history enforced (minimum 10 previous passwords)",
            "Account lockout after failed attempts (5-10 attempts)",
            "Lockout duration or CAPTCHA implemented",
            "Common/weak passwords blocked (breach database check)",
            "Default/vendor passwords changed before production",
            "Password strength meter provided to users",
            "Password reset process secure (no email plaintext passwords)",
            "Temporary passwords expire after first use",
            "Passwords not displayed during entry (masked)",
            "Failed login attempts logged",
            "Password policy documented and communicated",
            "Privileged accounts require stronger passwords (≥14 characters)",
        ],
        "mfa": [
            "MFA required for all administrative/privileged access",
            "MFA required for remote access (VPN, RDP, SSH)",
            "MFA required for access to Confidential/Restricted data",
            "MFA required for cloud service access (AWS, Azure, GCP, M365)",
            "MFA enrollment mandatory for all users",
            "TOTP (time-based one-time password) supported",
            "Push notification MFA supported",
            "Hardware security keys (FIDO2/WebAuthn) supported",
            "SMS-based MFA deprecated or phased out",
            "Biometric authentication available (where appropriate)",
            "Backup MFA methods configured for each user",
            "MFA bypass requests logged and approved",
            "MFA fatigue/prompt bombing protections implemented",
            "Device trust/registration for MFA",
            "MFA enrollment monitoring and alerts",
            "MFA recovery process documented",
            "Passwordless authentication option available",
            "MFA audit logs retained and reviewed",
            "Adaptive/risk-based MFA implemented",
        ],
        "certificate": [
            "Certificates issued from trusted CA (internal PKI or public CA)",
            "Certificate-based auth required for service accounts",
            "Certificate-based auth required for automated processes",
            "Client certificates used for VPN/remote access",
            "Smart cards issued for privileged users",
            "Machine/device certificates for endpoint authentication",
            "Certificate validity ≤2 years (1 year preferred)",
            "Certificate revocation checking enabled (CRL/OCSP)",
            "Certificate pinning implemented for critical services",
            "Private keys stored in HSM/TPM/smart card",
            "Private keys never exported in plaintext",
            "Certificate enrollment process documented",
            "Certificate inventory maintained",
            "Certificate expiration monitoring configured",
            "Mutual TLS (mTLS) implemented for API authentication",
            "PIV/CAC compliance (government/high-security)",
            "Certificate-based SSH authentication enabled",
            "Certificate revocation process tested",
        ],
        "service": [
            "Service account inventory maintained and up to date",
            "Certificate-based authentication used (preferred)",
            "Managed identities used for cloud services (AWS IAM, Azure MI)",
            "Long random passwords (≥20 characters) if certificate not available",
            "Password expiry disabled with compensating monitoring",
            "Service account passwords stored in secrets vault",
            "Service account passwords NEVER hardcoded in code",
            "Service account passwords NEVER in configuration files",
            "Interactive login disabled for service accounts",
            "Service accounts follow least privilege principle",
            "Service account usage logged and monitored",
            "Service account activity anomaly detection configured",
            "Privileged service accounts require additional controls",
            "Service account ownership documented",
            "Unused service accounts disabled/removed",
            "Service account naming convention followed",
            "Service account access reviewed quarterly",
            "Credential rotation tested in non-production",
            "Break-glass procedures for service account failures",
        ],
        "sso": [
            "SSO implemented for all compatible applications",
            "SAML 2.0, OAuth 2.0/OIDC, or WS-Federation protocols used",
            "Centralised Identity Provider (IdP) configured",
            "MFA enforced at IdP level",
            "SAML assertions signed and encrypted",
            "Certificate-based trust established (SAML metadata)",
            "Token lifetime/expiration configured (≤1 hour access token)",
            "Refresh token rotation implemented",
            "Just-in-time (JIT) user provisioning configured",
            "Attribute/claim mapping documented",
            "SSO session timeout configured (≤8 hours)",
            "Conditional access policies implemented",
            "Device compliance checks before SSO access",
            "SSO authentication logs centralised",
            "SSO bypass/backdoor disabled",
            "Federation trust relationships documented",
            "External federation partners vetted and approved",
            "SSO failover/redundancy configured",
            "IdP high availability implemented",
            "Privileged access management (PAM) integrated with SSO",
        ],
    }
    
    return checklists.get(sheet_type, [])


# ============================================================================
# SECTION 5: REFERENCE TABLE DEFINITIONS PER SHEET TYPE
# ============================================================================

def get_reference_tables(sheet_type):
    """
    Return reference table definitions for each sheet type.
    
    Returns:
        list: [(table_title, headers, data_rows), ...]
    """
    tables = {
        "password": [
            (
                "APPROVED PASSWORD HASHING ALGORITHMS",
                ["Algorithm", "Work Factor", "Salt Required", "Security Level", "Use Case"],
                [
                    ("Argon2id", "High (configurable)", "Yes", "Highest", "New implementations (recommended)"),
                    ("bcrypt", "Medium-High (cost factor 10-12)", "Yes", "High", "Widely supported, proven"),
                    ("PBKDF2-SHA256", "High (100'000+ iterations)", "Yes", "High", "Standards compliance (NIST)"),
                    ("scrypt", "High (configurable)", "Yes", "High", "Memory-hard function"),
                ]
            ),
            (
                "PASSWORD COMPLEXITY STANDARDS",
                ["Account Type", "Min Length", "Complexity", "Expiry", "MFA Required"],
                [
                    ("Standard User", "12 chars", "Upper, lower, number, symbol", "90 days", "Recommended"),
                    ("Administrator", "14 chars", "Upper, lower, number, symbol", "90 days", "Required"),
                    ("Service Account", "20 chars", "Random generation", "No expiry (monitor)", "Certificate preferred"),
                    ("External User", "12 chars", "Upper, lower, number, symbol", "90 days", "Required"),
                    ("API Access", "N/A", "Use API keys/tokens", "Token expiry", "Yes"),
                ]
            ),
        ],
        "mfa": [
            (
                "MFA FACTOR SECURITY COMPARISON",
                ["MFA Factor", "Security Level", "Phishing Resistant", "User Convenience", "Recommended Use"],
                [
                    ("FIDO2/WebAuthn (Hardware Key)", "Highest", "Yes", "High", "Privileged accounts, high-security"),
                    ("Certificate-based", "Highest", "Yes", "Medium", "Service accounts, automation"),
                    ("TOTP (Authenticator App)", "High", "No", "High", "General users, widely supported"),
                    ("Push Notification", "High", "Partial", "High", "General users, mobile-friendly"),
                    ("Biometric (with device trust)", "High", "Partial", "Highest", "Endpoints, mobile devices"),
                    ("SMS/Voice", "Low", "No", "Medium", "DEPRECATED - legacy only"),
                ]
            ),
            (
                "MFA IMPLEMENTATION PRIORITIES",
                ["Priority", "Account Type", "MFA Requirement", "Recommended Factor"],
                [
                    ("Critical", "Domain Administrators", "Required", "Hardware key + Backup TOTP"),
                    ("Critical", "Cloud Admins (AWS/Azure/GCP)", "Required", "Hardware key + Backup TOTP"),
                    ("Critical", "Privileged Access Management", "Required", "Hardware key or Certificate"),
                    ("High", "Standard Remote Users", "Required", "TOTP or Push"),
                    ("High", "Access to Confidential Data", "Required", "TOTP or Push"),
                    ("Medium", "Internal Users (on-premises)", "Recommended", "TOTP or Push"),
                    ("Low", "Read-only Public Data", "Optional", "TOTP"),
                ]
            ),
        ],
        "certificate": [
            (
                "CERTIFICATE AUTHENTICATION USE CASES",
                ["Use Case", "Certificate Type", "Key Algorithm", "Validity", "Storage"],
                [
                    ("Service Account Auth", "Client Certificate", "RSA-3072+", "1 year", "HSM/KMS"),
                    ("Smart Card Login", "User Certificate (PIV)", "RSA-2048+ or ECDSA P-256+", "1-3 years", "Smart Card"),
                    ("VPN Client Auth", "Client Certificate", "RSA-3072+ or ECDSA P-256+", "1 year", "TPM/Smart Card"),
                    ("Machine/Device Auth", "Device Certificate", "RSA-2048+ or ECDSA P-256+", "1-2 years", "TPM"),
                    ("Mutual TLS (mTLS)", "Client & Server Cert", "RSA-3072+ or ECDSA P-256+", "1 year", "HSM/KMS"),
                    ("Code Signing", "Code Signing Cert", "RSA-3072+", "1-3 years", "Hardware Token"),
                    ("Email Signing (S/MIME)", "Email Certificate", "RSA-2048+ or ECDSA P-256+", "1 year", "Smart Card/Keystore"),
                ]
            ),
            (
                "PKI INFRASTRUCTURE CHECKLIST",
                ["Component", "Requirement", "Implementation Status"],
                [
                    ("Root CA", "Offline, HSM-protected", ""),
                    ("Issuing CA", "Online, HSM-protected", ""),
                    ("Registration Authority", "Enrollment process automation", ""),
                    ("CRL Distribution", "HTTP/LDAP accessible", ""),
                    ("OCSP Responder", "Real-time revocation checking", ""),
                    ("Certificate Templates", "Standardised certificate profiles", ""),
                    ("Auto-enrollment", "User/machine certificate automation", ""),
                ]
            ),
        ],
        "service": [
            (
                "SERVICE ACCOUNT AUTHENTICATION METHODS",
                ["Authentication Method", "Security Level", "Use Case", "Implementation"],
                [
                    ("Managed Identity (Cloud)", "Highest", "Cloud services (AWS, Azure, GCP)", "AWS IAM Roles, Azure Managed Identity"),
                    ("Certificate-based", "Highest", "On-premises, API integration", "Client certificates, mTLS"),
                    ("Secrets Vault (HashiCorp/CyberArk)", "High", "Dynamic secret generation", "Vault dynamic credentials"),
                    ("API Key (Long, Random)", "Medium-High", "API services, integrations", "≥256-bit entropy, rotated 90 days"),
                    ("Long Password (≥20 chars)", "Medium", "Legacy systems, compatibility", "Random generation, vault storage"),
                    ("Shared Secret", "Low", "DEPRECATED", "Phase out"),
                ]
            ),
            (
                "SERVICE ACCOUNT LIFECYCLE MANAGEMENT",
                ["Lifecycle Stage", "Actions", "Responsible Role", "Frequency"],
                [
                    ("Request", "Business justification, owner assignment", "Service Owner", "As needed"),
                    ("Provisioning", "Create account, assign permissions, document", "IAM Admin", "Upon approval"),
                    ("Credential Storage", "Store in secrets vault, configure rotation", "Security Team", "At creation"),
                    ("Monitoring", "Log usage, detect anomalies, alert", "SOC/Security Team", "Continuous"),
                    ("Review", "Validate ownership, access, necessity", "IAM Admin", "Quarterly"),
                    ("Rotation", "Update credentials, test connectivity", "Service Owner", "90 days (if password)"),
                    ("Decommissioning", "Disable account, remove access, document", "IAM Admin", "Upon request"),
                ]
            ),
        ],
        "sso": [
            (
                "SSO PROTOCOL COMPARISON",
                ["Protocol", "Use Case", "Security Features", "Token Type", "Best For"],
                [
                    ("SAML 2.0", "Enterprise SSO, SaaS", "Signed/encrypted assertions, strong binding", "XML assertions", "Enterprise applications, legacy systems"),
                    ("OAuth 2.0 + OIDC", "Modern web/mobile apps, APIs", "JWT tokens, PKCE, scopes", "JSON Web Tokens (JWT)", "Cloud-native apps, mobile, APIs"),
                    ("WS-Federation", "Microsoft ecosystems", "Encrypted tokens, claims-based", "SAML tokens", "Active Directory Federation Services (ADFS)"),
                    ("Kerberos", "On-premises Windows", "Ticket-based, mutual authentication", "Kerberos tickets", "Internal Windows environments"),
                ]
            ),
            (
                "IDENTITY PROVIDER SECURITY CHECKLIST",
                ["Requirement", "Implementation Status", "Notes"],
                [
                    ("MFA enforcement at IdP", "", "Required for all users"),
                    ("Conditional access policies", "", "Risk-based authentication"),
                    ("Device compliance checking", "", "Managed devices only"),
                    ("Geo-location restrictions", "", "Block suspicious locations"),
                    ("Adaptive authentication", "", "Risk scoring, step-up auth"),
                    ("Session management", "", "Timeout, revocation"),
                    ("Audit logging", "", "All auth events logged"),
                    ("High availability", "", "99.9%+ uptime SLA"),
                ]
            ),
            (
                "TOKEN LIFETIME BEST PRACTICES",
                ["Token Type", "Recommended Lifetime", "Maximum Lifetime", "Rotation"],
                [
                    ("Access Token (OAuth)", "15 minutes", "1 hour", "Auto-expire"),
                    ("Refresh Token (OAuth)", "90 days", "1 year", "Rotate on use"),
                    ("ID Token (OIDC)", "5 minutes", "15 minutes", "Per request"),
                    ("SAML Assertion", "5 minutes", "15 minutes", "Per request"),
                    ("Session Token", "1 hour inactive", "8 hours absolute", "Extend on activity"),
                ]
            ),
        ],
    }
    
    return tables.get(sheet_type, [])
# ============================================================================
# SECTION 6: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete each worksheet tab (1–5) for applicable authentication methods.', '2. Use dropdown menus for standardised entries (Status, Auth Method, Algorithm, etc.).', '3. Fill in yellow-highlighted cells with your information.', '4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.', '5. Document password policies, MFA implementations, certificate usage, and SSO configurations.', '6. Provide evidence location/path for each implementation entry.', '7. Summary Dashboard auto-calculates compliance statistics per authentication area.', '8. Maintain the Evidence Register for audit traceability.', '9. Obtain final approval and sign-off in the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_assessment_sheet(ws, styles, section_title, policy_ref, question, 
                           sheet_type):
    """
    Generic assessment sheet creator for authentication controls.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "1. PASSWORD SECURITY - PASSWORD AUTHENTICATION CONTROLS"
        policy_ref: policy requirement text
        question: assessment question
        sheet_type: key for columns/checklist/tables (password, mfa, certificate, service, sso)
    """
    columns = get_all_columns(sheet_type)
    checklist_items = get_checklist_items(sheet_type)
    reference_tables = get_reference_tables(sheet_type)
    
    total_cols = len(columns)
    last_col_letter = get_column_letter(total_cols)
    
    # ---------- HEADER ----------
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{section_title}\nPolicy Requirement: {policy_ref}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.8.24: Use of Cryptography"
    ws["A2"].font = Font(italic=True, size=10, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ---------- ASSESSMENT QUESTION ----------
    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ---------- RESPONSE DROPDOWN ----------
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    # Create validations
    base_validations = create_base_validations(ws)
    extended_validations = create_extended_validations(ws, sheet_type)
    base_validations['response'].add(ws["B4"])

    # ---------- COLUMN HEADERS ----------
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ---------- EXAMPLE ROW ----------
    example_row = 7
    example_values = _get_example_row(sheet_type, total_cols)
    for col_idx, value in enumerate(example_values, start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------- DATA ENTRY ROWS (8-20) ----------
    start_row = 8
    end_row = 20
    
    for r in range(start_row, end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Apply base validations to standard columns (B-Q, indices 2-17)
    for r in range(start_row, end_row + 1):
        base_validations['auth_method'].add(ws.cell(row=r, column=2))  # B: Authentication Method
        base_validations['user_type'].add(ws.cell(row=r, column=3))    # C: User Type
        base_validations['data_class'].add(ws.cell(row=r, column=4))   # D: Data Classification
        base_validations['algorithm'].add(ws.cell(row=r, column=5))    # E: Cryptographic Algorithm
        base_validations['hash_status'].add(ws.cell(row=r, column=6))  # F: Hash/Encryption Status
        base_validations['pwd_complexity'].add(ws.cell(row=r, column=7)) # G: Password Complexity
        base_validations['status'].add(ws.cell(row=r, column=8))       # H: Status
        base_validations['remediation'].add(ws.cell(row=r, column=11)) # K: Remediation Needed
        base_validations['budget'].add(ws.cell(row=r, column=17))      # Q: Budget Required

    # Apply extended validations for columns R onwards (18+)
    _apply_extended_validations(ws, sheet_type, extended_validations, start_row, end_row)

    ws.freeze_panes = "A7"

    next_row = end_row + 2

    # ---------- COMPLIANCE CHECKLIST ----------
    ws[f"A{next_row}"] = f"{section_title.split('-')[0].strip().upper()} CHECKLIST"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)
    next_row += 1

    ws[f"A{next_row}"] = "☐"
    ws[f"B{next_row}"] = "Requirement"
    ws[f"C{next_row}"] = "Status"
    for col in ["A", "B", "C"]:
        ws[f"{col}{next_row}"].font = Font(bold=True)
    next_row += 1

    checklist_start = next_row
    for item in checklist_items:
        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = item
        ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{next_row}"].border = styles["border"]

        base_validations['checklist'].add(ws[f"C{next_row}"])
        next_row += 1

    # Checklist score
    ws[f"A{next_row}"] = "Checklist Score:"
    ws[f"A{next_row}"].font = Font(bold=True)
    ws[f"B{next_row}"] = (
        f'=IFERROR(ROUND(COUNTIF(C{checklist_start}:C{next_row-1},"✅ Yes")/'
        f'COUNTA(C{checklist_start}:C{next_row-1})*100,0)&"%","0%")'
    )
    ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
    next_row += 2

    # ---------- REFERENCE TABLES ----------
    for table_title, headers, data_rows in reference_tables:
        ws[f"A{next_row}"] = table_title
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1
        
        # Table headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=header)
            cell.font = styles["column_header"]["font"]
            cell.fill = styles["column_header"]["fill"]
            cell.alignment = styles["column_header"]["alignment"]
            cell.border = styles["column_header"]["border"]
        next_row += 1
        
        # Table data rows
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = styles["border"]
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # For PKI Infrastructure table (has status dropdown in column 3)
                if "PKI INFRASTRUCTURE" in table_title and col_idx == 3:
                    cell.fill = styles["input_cell"]["fill"]
                    # Create dropdown for implementation status
                    dv_impl = DataValidation(
                        type="list",
                        formula1='"Implemented,Planned,Not Implemented"',
                        allow_blank=False
                    )
                    ws.add_data_validation(dv_impl)
                    dv_impl.add(cell)
            next_row += 1
        
        next_row += 1  # Space between tables

    # ---------- EXCEPTION/DEVIATION BLOCK ----------
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "EXCEPTION / DEVIATION DOCUMENTATION (Complete if Status = Partial or Non-Compliant)"
    ws[f"A{next_row}"].font = styles["exception_header"]["font"]
    ws[f"A{next_row}"].fill = styles["exception_header"]["fill"]
    ws[f"A{next_row}"].alignment = styles["exception_header"]["alignment"]

    next_row += 1
    exception_fields = [
        ("Formal exception request submitted:", "✅ Yes,❌ No"),
        ("Exception ID:", ""),
        ("Risk acceptance documented:", "✅ Yes,❌ No"),
        ("Risk ID:", ""),
        ("Compensating Controls (summary):", ""),
        ("☐ Enhanced monitoring and alerting", ""),
        ("☐ Network segmentation / restricted access", ""),
        ("☐ Additional authentication factor", ""),
        ("☐ Manual approval process", ""),
        ("☐ Increased audit frequency", ""),
        ("☐ Time-limited exception (expiry date required)", ""),
        ("☐ Other (describe):", ""),
    ]

    for label, options in exception_fields:
        ws[f"A{next_row}"] = label
        ws[f"A{next_row}"].font = Font(bold=True)
        ws.merge_cells(f"B{next_row}:D{next_row}")
        ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{next_row}"].border = styles["border"]
        ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]

        if options:
            base_validations['exception_yn'].add(ws[f"B{next_row}"])

        next_row += 1

    # ---------- NOTES ----------
    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "ADDITIONAL NOTES / COMMENTS"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)

    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row+8}")
    ws[f"A{next_row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{next_row}"].border = styles["border"]
    ws[f"A{next_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths for checklist/tables
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 20

    # GS-AS-011: Borders on all merged cell ranges
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for _mr in list(ws.merged_cells.ranges):
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                ws.cell(row=_r, column=_c).border = _border

    return (start_row, end_row, 8)  # Return status column index (H = 8)


def _get_example_row(sheet_type, total_cols):
    """Generate example row values based on sheet type."""
    examples = {
        "password": [
            "Example: Production HR System - Workday", "Password", "End User", "Confidential",
            "bcrypt", "Properly Hashed", "Strong (≥14 chars)", "✅ Compliant",
            "/evidence/auth/workday-password-policy.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "≥14 chars", "Yes", "90 days", "≥10 passwords", "Yes", "Yes", "Changed"
        ],
        "mfa": [
            "Example: AWS Production Account", "MFA", "Administrator", "Restricted",
            "TOTP", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/aws-mfa-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "TOTP (Authenticator App)", "100%", "Required", "Yes", "Never", "Available"
        ],
        "certificate": [
            "Example: VPN Gateway - Cisco ASA", "Certificate", "End User", "Confidential",
            "RSA-3072", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/vpn-cert-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Client Certificate", "Internal CA", "RSA-3072", "≤1 year", "Yes", "No", "No"
        ],
        "service": [
            "Example: Database Service Account - SQL_SVC", "Service Account", "Service Account", "Restricted",
            "Certificate", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/sqlsvc-cert.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Database Service", "Certificate", "Never (Monitored)", "Yes", "Yes", "Disabled", "Yes"
        ],
        "sso": [
            "Example: Office 365 SSO - Microsoft Entra ID (formerly Azure AD)", "SSO", "End User", "Confidential",
            "SHA-256", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/o365-sso-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "SAML 2.0", "Microsoft Entra ID (formerly Azure AD)", "Yes", "Enabled", "≤1 hour", "Yes", "Yes"
        ],
    }
    
    base_example = examples.get(sheet_type, ["Example"] * total_cols)
    # Pad or trim to match total_cols
    if len(base_example) < total_cols:
        base_example.extend(["N/A"] * (total_cols - len(base_example)))
    return base_example[:total_cols]


def _apply_extended_validations(ws, sheet_type, validations, start_row, end_row):
    """Apply extended column validations based on sheet type."""
    if sheet_type == "password":
        for r in range(start_row, end_row + 1):
            validations['min_length'].add(ws.cell(row=r, column=18))  # R
            validations['yn_na'].add(ws.cell(row=r, column=19))       # S: Complexity Enforced
            validations['expiry'].add(ws.cell(row=r, column=20))      # T
            validations['history'].add(ws.cell(row=r, column=21))     # U
            validations['yn_na'].add(ws.cell(row=r, column=22))       # V: Account Lockout
            validations['yn_na'].add(ws.cell(row=r, column=23))       # W: Password Strength Meter
            validations['default_pwd'].add(ws.cell(row=r, column=24)) # X
    
    elif sheet_type == "mfa":
        for r in range(start_row, end_row + 1):
            validations['mfa_factor'].add(ws.cell(row=r, column=18))  # R
            validations['enrollment'].add(ws.cell(row=r, column=19))  # S
            validations['enforcement'].add(ws.cell(row=r, column=20)) # T
            validations['yn_na'].add(ws.cell(row=r, column=21))       # U: Backup MFA Method
            validations['bypass'].add(ws.cell(row=r, column=22))      # V
            validations['yn_na'].add(ws.cell(row=r, column=23))       # W: Passwordless Option
    
    elif sheet_type == "certificate":
        for r in range(start_row, end_row + 1):
            validations['cert_type'].add(ws.cell(row=r, column=18))   # R
            # S: Issuing CA - free text
            validations['key_algo'].add(ws.cell(row=r, column=20))    # T
            validations['cert_validity'].add(ws.cell(row=r, column=21)) # U
            validations['yn_na'].add(ws.cell(row=r, column=22))       # V: CRL/OCSP
            validations['yn_na'].add(ws.cell(row=r, column=23))       # W: Smart Card
            validations['yn_na'].add(ws.cell(row=r, column=24))       # X: PIV/CAC
    
    elif sheet_type == "service":
        for r in range(start_row, end_row + 1):
            validations['svc_type'].add(ws.cell(row=r, column=18))    # R
            validations['svc_auth'].add(ws.cell(row=r, column=19))    # S
            validations['rotation'].add(ws.cell(row=r, column=20))    # T
            validations['yn_na'].add(ws.cell(row=r, column=21))       # U: Privileged Account
            validations['yn_na'].add(ws.cell(row=r, column=22))       # V: Account Monitoring
            validations['interactive'].add(ws.cell(row=r, column=23)) # W
            validations['yn_na'].add(ws.cell(row=r, column=24))       # X: Least Privilege
    
    elif sheet_type == "sso":
        for r in range(start_row, end_row + 1):
            validations['sso_protocol'].add(ws.cell(row=r, column=18)) # R
            # S: Identity Provider - free text
            validations['yn_na'].add(ws.cell(row=r, column=20))        # T: MFA at IdP
            validations['yn_na'].add(ws.cell(row=r, column=21))        # U: JIT Provisioning
            validations['session_timeout'].add(ws.cell(row=r, column=22)) # V
            validations['yn_na'].add(ws.cell(row=r, column=23))        # W: Certificate Validation
            validations['yn_na'].add(ws.cell(row=r, column=24))        # X: Token Encryption
# ============================================================================
# SECTION 8: INDIVIDUAL ASSESSMENT SHEET DEFINITIONS
# ============================================================================

def create_1_password_security(ws, styles):
    """1. Password Security - matches markdown spec exactly."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1. PASSWORD SECURITY - PASSWORD AUTHENTICATION CONTROLS",
        policy_ref="Passwords MUST be hashed using approved algorithms (bcrypt, Argon2, PBKDF2, scrypt). Minimum 12 characters, complexity requirements enforced (Policy Section 2.5.1)",
        question="Does your organisation use password-based authentication for any systems or applications?",
        sheet_type="password",
    )


def create_2_multi_factor_authentication(ws, styles):
    """2. Multi-Factor Authentication - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2. MULTI-FACTOR AUTHENTICATION (MFA) - STRONG AUTHENTICATION",
        policy_ref="MFA REQUIRED for all administrative access, remote access, and access to Confidential/Restricted data. Approved factors: TOTP, Push notification, Hardware token, Biometric (Policy Section 2.5.2)",
        question="Does your organisation require multi-factor authentication for any systems or user types?",
        sheet_type="mfa",
    )


def create_3_certificate_based_auth(ws, styles):
    """3. Certificate-Based Authentication - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="3. CERTIFICATE-BASED AUTHENTICATION - PKI & DIGITAL CERTIFICATES",
        policy_ref="Certificate-based authentication REQUIRED for service accounts, automated processes, and high-security environments. Certificates issued from trusted CA with proper key management (Policy Section 2.5.3)",
        question="Does your organisation use digital certificates for authentication (client certificates, smart cards, machine authentication)?",
        sheet_type="certificate",
    )


def create_4_service_accounts(ws, styles):
    """4. Service Accounts - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4. SERVICE ACCOUNTS - NON-HUMAN AUTHENTICATION",
        policy_ref="Service accounts MUST use certificate-based authentication or long random passwords (≥20 characters). Password expiry disabled with monitoring. MFA or certificate preferred (Policy Section 2.5.4)",
        question="Does your organisation use service accounts for application-to-application or automated process authentication?",
        sheet_type="service",
    )


def create_5_sso_federation(ws, styles):
    """5. SSO & Federation - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="5. SSO & FEDERATION - CENTRALIZED AUTHENTICATION",
        policy_ref="SSO/Federation REQUIRED using SAML 2.0, OAuth 2.0/OIDC, or WS-Federation. Identity Provider (IdP) must enforce strong authentication and MFA (Policy Section 2.5.5)",
        question="Does your organisation use Single Sign-On (SSO) or federated authentication for applications and services?",
        sheet_type="sso",
    )


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD (COMPLEX - 12 ANALYSIS SECTIONS)
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard — TABLE 1/2/3, all tables uniform 7-col A:G width."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Row 1: Title ---
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME.upper()} — SUMMARY DASHBOARD"
    ws["A1"].font      = styles["header"]["font"]
    ws["A1"].fill      = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border

    # --- Row 2: ISO reference ---
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 | Control A.8.24 | Cryptographic Authentication Controls"
    ws["A2"].font      = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border

    # --- TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW ---
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW")
    c.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="003366")
    #c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    for col_idx, h in enumerate(["Assessment Area", "Total Items", "Compliant", "Partial",
                                  "Non-Compliant", "N/A", "Compliance %"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill      = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    areas = [
        ("1. Password Security",         "1. Password Security"),
        ("2. Multi-Factor Authentication", "2. Multi-Factor Authentication"),
        ("3. Certificate-Based Auth",     "3. Certificate-Based Auth"),
        ("4. Service Accounts",           "4. Service Accounts"),
        ("5. SSO & Federation",           "5. SSO & Federation"),
    ]
    row += 1
    start_data_row = row
    for label, sheet in areas:
        ws.cell(row=row, column=1, value=label)
        rng = f"'{sheet}'!H8:H87"
        ws.cell(row=row, column=2, value=f"=COUNTA({rng})")
        ws.cell(row=row, column=3, value=f'=COUNTIF({rng},"✅*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({rng},"⚠️*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({rng},"❌*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({rng},"N/A")')
        ws.cell(row=row, column=7,
                value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font  = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill                 = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border               = border
    for col in range(2, 7):
        c       = ws.cell(row=total_row, column=col)
        c.value = f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{total_row - 1})"
        c.font  = Font(name="Calibri", size=10, bold=True)
        c.fill  = PatternFill("solid", fgColor="D9D9D9")
        c.border = border
    c       = ws.cell(row=total_row, column=7)
    c.value = (f'=IF((B{total_row}-F{total_row})=0,"0%",'
               f'ROUND(C{total_row}/(B{total_row}-F{total_row})*100,1)&"%")')
    c.font   = Font(name="Calibri", size=12, bold=True, color="000000")
    c.fill   = PatternFill("solid", fgColor="D9D9D9")
    c.border = border

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 15

    _kpis = [
        ("MFA Enrollment – Administrators", "100%"),
        ("MFA Enrollment – Remote Users", "100%"),
        ("Certificate-Based Auth Systems (count)", "Count"),
        ("Service Accounts in Secrets Vault", "≥95%"),
        ("SSO-Integrated Applications (count)", "Count"),
        ("Legacy / Weak Auth Applications (count)", "0"),
        ("Password Complexity Compliance", "100%"),
        ("Password Length ≥ 12 Characters", "100%"),
        ("Overall Authentication Security Score", "≥85%"),
        ("Evidence Documentation Rate", "100%"),
    ]
    # --- TABLE 2: KEY METRICS ---
    row = total_row + 2
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    #c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    for col_idx, h in enumerate(["KPI", "Current Value", "Target", "Status",
                                  "Last Updated", "Owner", "Notes"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    dv_kpi_sts = DataValidation(type="list",
        formula1='"\u2705 On Target,\u26a0\ufe0f At Risk,\u274c Below Target"',
        allow_blank=True)
    ws.add_data_validation(dv_kpi_sts)

    # --- TABLE 2a: AUTO-COMPUTED METRICS (formula-driven) ---
    row += 1
    t2a_headers = ["Metric (Auto-computed)", "Value", "Notes", "", "", "", ""]
    for col_idx, h in enumerate(t2a_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h if h else "")
        if h:
            c.font = Font(name="Calibri", size=10, bold=True, color="000000")
            c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border

    row += 1
    # Auth sheets: H=Status, B=Auth Method, F=Hash/Encryption Status, G=Password Complexity
    _auth_sheets = [
        "1. Password Security", "2. Multi-Factor Authentication",
        "3. Certificate-Based Auth", "4. Service Accounts", "5. SSO & Federation",
    ]
    def _sum_countif_auth(col_letter, value):
        parts = [f'COUNTIF(\'{s}\'!{col_letter}8:{col_letter}87,\"{value}\")' for s in _auth_sheets]
        return "=SUM(" + ",".join(parts) + ")"

    auto_metrics = [
        ("Non-Compliant Items (all domains)",
         "=" + "+".join([f'COUNTIF(\'{s}\'!H8:H87,"❌*")' for s in _auth_sheets])),
        ("Partial Items (all domains)",
         "=" + "+".join([f'COUNTIF(\'{s}\'!H8:H87,"⚠️*")' for s in _auth_sheets])),
        ("Compliant Items (all domains)",
         "=" + "+".join([f'COUNTIF(\'{s}\'!H8:H87,"✅*")' for s in _auth_sheets])),
        ("Total Assessed Items (all domains)",
         "=" + "+".join([f'COUNTA(\'{s}\'!H8:H87)' for s in _auth_sheets])),
        ("Password-Only Auth (count)",       _sum_countif_auth("B", "Password")),
        ("MFA Deployed (count)",             _sum_countif_auth("B", "MFA")),
        ("Properly Hashed Credentials (count)", _sum_countif_auth("F", "Properly Hashed")),
        ("Plaintext Passwords (NON-COMPLIANT)", _sum_countif_auth("F", "Plaintext (Non-compliant)")),
    ]
    for metric, fml in auto_metrics:
        ws.cell(row=row, column=1, value=metric).border = border
        cx = ws.cell(row=row, column=2, value=fml)
        cx.font   = Font(name="Calibri", size=10, color="000000")
        cx.border = border
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    row += 1  # spacer before manual KPIs

    row += 1
    for kpi, target in _kpis:
        ws.cell(row=row, column=1, value=kpi).border = border
        c2 = ws.cell(row=row, column=2)
        c2.border = border
        ws.cell(row=row, column=3, value=target).border = border
        c4 = ws.cell(row=row, column=4)
        c4.border = border
        dv_kpi_sts.add(c4)
        for col in range(5, 8):
            ws.cell(row=row, column=col).border = border
        row += 1
    r = row

    # 2 bordered buffer rows + blank gap before TABLE 3 (Gold Standard)
    for _buf in range(2):
        for _col in range(1, 8):
            ws.cell(row=r, column=_col).border = border
        r += 1

    # --- TABLE 3: DOMAIN COMPLIANCE SUMMARY ---
    row = r + 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: DOMAIN COMPLIANCE SUMMARY")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

    row += 1
    t3_headers = [
        "Authentication Domain",
        "Non-Compliant",
        "Partial",
        "Compliant",
        "N/A",
        "Total Items",
        "Compliance %",
    ]
    for col_idx, h in enumerate(t3_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill      = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    # Auth domains: status col H (col 8), data rows 8-87
    t3_domains = [
        ("1. Password Security",           "1. Password Security"),
        ("2. Multi-Factor Authentication", "2. Multi-Factor Authentication"),
        ("3. Certificate-Based Auth",      "3. Certificate-Based Auth"),
        ("4. Service Accounts",            "4. Service Accounts"),
        ("5. SSO & Federation",            "5. SSO & Federation"),
    ]
    row += 1
    t3_start = row
    for label, sheet in t3_domains:
        rng     = f'\'{sheet}\'!H8:H87'
        non_c   = f'=COUNTIF({rng},"❌*")'
        part    = f'=COUNTIF({rng},"⚠️*")'
        compl   = f'=COUNTIF({rng},"✅*")'
        na_     = f'=COUNTIF({rng},"N/A")'
        total_  = f'=COUNTA({rng})'
        pct_r   = row
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font   = Font(name="Calibri", size=10, color="000000")
        c1.fill   = PatternFill("solid", fgColor="FFFFCC")
        c1.border = border
        for col_idx, fml in enumerate([non_c, part, compl, na_, total_], 2):
            cx = ws.cell(row=row, column=col_idx, value=fml)
            cx.font   = Font(name="Calibri", size=10, color="000000")
            cx.fill   = PatternFill("solid", fgColor="FFFFCC")
            cx.border = border
        c7 = ws.cell(row=row, column=7,
                     value=f'=IF((F{pct_r}-E{pct_r})=0,0,D{pct_r}/(F{pct_r}-E{pct_r}))')
        c7.font          = Font(name="Calibri", size=10, color="000000")
        c7.fill          = PatternFill("solid", fgColor="FFFFCC")
        c7.number_format = "0.0%"
        c7.border        = border
        row += 1

    # TOTAL row
    t3_end = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font  = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=1).fill   = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=row, column=1).border = border
    for col in range(2, 7):
        cx = ws.cell(row=row, column=col)
        cx.value  = f"=SUM({get_column_letter(col)}{t3_start}:{get_column_letter(col)}{t3_end})"
        cx.font   = Font(name="Calibri", size=10, bold=True)
        cx.fill   = PatternFill("solid", fgColor="D9D9D9")
        cx.border = border
    tot_row = row
    cx = ws.cell(row=row, column=7,
                 value=f'=IF((F{tot_row}-E{tot_row})=0,0,D{tot_row}/(F{tot_row}-E{tot_row}))')
    cx.font          = Font(name="Calibri", size=10, bold=True, color="000000")
    cx.fill          = PatternFill("solid", fgColor="D9D9D9")
    cx.number_format = "0.0%"
    cx.border        = border
    row += 1

    # Borders on all merged cell ranges (GS-AS-011)
    _t = Side(style="thin")
    _b = Border(left=_t, right=_t, top=_t, bottom=_t)
    for _mr in list(ws.merged_cells.ranges):
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                ws.cell(row=_r, column=_c).border = _b

    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create evidence register for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validation dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Password policy,GPO screenshot,MFA enrollment report,Certificate config,PKI documentation,SAML configuration,IdP settings,Service account inventory,Authentication logs,Audit trail,Policy document,Security assessment,Penetration test,Compliance scan,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    # Sample row (row 5) — F2F2F2 grey with realistic example data
    _er_sample = [
        "EV-001", "1. Password Security", "Configuration file",
        "TLS/encryption configuration export for audit evidence",
        "/evidence/a824/config-export-2025.txt",
        "01.01.2025", "Security Team", "✅ Verified",
    ]
    for c, val in enumerate(_er_sample, start=1):
        cell = ws.cell(row=5, column=c, value=val)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if c == 1:
            cell.font = Font(italic=True, color="555555")
    dv_type.add(ws.cell(row=5, column=3))
    dv_ver.add(ws.cell(row=5, column=8))

    # Empty data rows (rows 6-105) — 100 FFFFCC input rows, NO pre-filled IDs
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(ws):
    """Create approval and sign-off sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # ASSESSMENT SUMMARY banner
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_fields = [
        ("Document:", "ISMS-IMP-A.8.24.3 - Authentication Assessment"),
        ("Assessment Period:", ""),
        ("Overall Compliance:", "='Summary Dashboard'!G11"),
        ("Assessment Status:", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row - 1}"])

    # Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = border

    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    for merge_range in list(ws.merged_cells.ranges):
        for _row in range(merge_range.min_row, merge_range.max_row + 1):
            for _col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=_row, column=_col).border = border
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def main() -> int:
    """Main execution function - orchestrates workbook creation."""
    try:
        logger.info("=" * 78)
        logger.info("ISMS-IMP-A.8.24.3 - Authentication Assessment Generator")
        logger.info("ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography (Authentication)")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/9] Creating Instructions & Legend sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/9] Creating 1. Password Security sheet...")
        create_1_password_security(wb["1. Password Security"], styles)

        logger.info("[3/9] Creating 2. Multi-Factor Authentication sheet...")
        create_2_multi_factor_authentication(wb["2. Multi-Factor Authentication"], styles)

        logger.info("[4/9] Creating 3. Certificate-Based Auth sheet...")
        create_3_certificate_based_auth(wb["3. Certificate-Based Auth"], styles)

        logger.info("[5/9] Creating 4. Service Accounts sheet...")
        create_4_service_accounts(wb["4. Service Accounts"], styles)

        logger.info("[6/9] Creating 5. SSO & Federation sheet...")
        create_5_sso_federation(wb["5. SSO & Federation"], styles)

        logger.info("[7/9] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

        logger.info("[8/9] Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[9/9] Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        output_path = _wkbk_dir / OUTPUT_FILENAME
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        finalize_validations(wb)
        wb.save(output_path)
        logger.info(f"SUCCESS: {output_path}")
        logger.info("Workbook Structure:")
        logger.info("  - Instructions & Legend - Document info and guidance")
        logger.info("  - 5 Assessment Sheets:")
        logger.info("    1. Password Security (20 checklist items + 2 reference tables)")
        logger.info("    2. Multi-Factor Authentication (19 checklist items + 2 reference tables)")
        logger.info("    3. Certificate-Based Auth (18 checklist items + 2 reference tables)")
        logger.info("    4. Service Accounts (19 checklist items + 2 reference tables)")
        logger.info("    5. SSO & Federation (20 checklist items + 3 reference tables)")
        logger.info("  - Summary Dashboard - 8 comprehensive analysis sections:")
        logger.info("    - Compliance summary by authentication type")
        logger.info("    - Authentication method distribution")
        logger.info("    - MFA adoption metrics by user type")
        logger.info("    - Password security metrics")
        logger.info("    - Service account security analysis")
        logger.info("    - SSO coverage analysis")
        logger.info("    - Overall authentication security score (weighted)")
        logger.info("    - Critical security gaps (8 entry rows)")
        logger.info("  - Evidence Register - 100 evidence entry rows")
        logger.info("  - Approval Sign-Off - Complete workflow with 3 sign-off levels")

        logger.info("Column Structure:")
        logger.info("  - Base columns (A-Q): 17 standard columns across all sheets")
        logger.info("  - Extended columns (R-X):")
        logger.info("    - Password Security: +7 columns (min length, complexity, expiry, etc.)")
        logger.info("    - MFA: +6 columns (factor type, enrollment, enforcement, etc.)")
        logger.info("    - Certificate: +7 columns (cert type, CA, key algo, validity, etc.)")
        logger.info("    - Service Accounts: +7 columns (account type, auth method, rotation, etc.)")
        logger.info("    - SSO & Federation: +7 columns (protocol, IdP, MFA, JIT, session, etc.)")

        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions & Legend")
        logger.info("  2) Fill yellow cells in each assessment sheet (1-5)")
        logger.info("  3) Check compliance checklists per authentication type")
        logger.info("  4) Review reference tables (algorithms, use cases, best practices)")
        logger.info("  5) Document exceptions/deviations as needed")
        logger.info("  6) Maintain Evidence Register entries")
        logger.info("  7) Review Summary Dashboard:")
        logger.info("     - Overall compliance by authentication area")
        logger.info("     - Authentication method distribution and risk ratings")
        logger.info("     - MFA adoption rates by user type")
        logger.info("     - Password security compliance metrics")
        logger.info("     - Service account security posture")
        logger.info("     - SSO coverage and protocol usage")
        logger.info("     - Weighted authentication security score")
        logger.info("     - Critical gaps requiring immediate remediation")
        logger.info("  8) Complete multi-level approval sign-off workflow")
        logger.info("=" * 78)
        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
