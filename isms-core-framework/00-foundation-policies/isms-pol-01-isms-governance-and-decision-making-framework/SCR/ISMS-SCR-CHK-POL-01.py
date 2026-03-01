#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-CHK-POL-01 - ISMS Governance Framework Compliance Assessment
================================================================================

ISO/IEC 27001:2022 Clause 5.3: Roles, responsibilities and authorities
Governance Framework Assessment

This script generates a comprehensive Excel assessment workbook for evaluating
governance process compliance with ISMS-POL-01 (ISMS Governance and Decision-
Making Framework).

Assessment Domains:
1. Authority Boundaries (Section 2)
2. Applicability Decisions (Section 3)
3. Exception Handling (Section 4)
4. Change Management (Section 5)
5. Governance Review (Section 6)

Total Requirements: 20 (GOV-01 through GOV-20)
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-CHK-POL-01"
WORKBOOK_NAME = "Governance Framework Compliance Assessment"
POLICY_ID = "ISMS-POL-01"
POLICY_NAME = "ISMS Governance and Decision-Making Framework"
POLICY_REF = f"{POLICY_ID}: {POLICY_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output path — save to WKBK/ folder (SRC-023)
from pathlib import Path as _Path
_wkbk_dir = _Path(__file__).parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
OUTPUT_FILENAME = str(_wkbk_dir / f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx")

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = "\u2705"
WARNING = "\u26a0\ufe0f"
XMARK = "\u274c"
DASH = "\u2014"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def finalize_validations(ws, validations):
    """Batch-apply DataValidation objects to worksheet."""
    for dv in validations:
        ws.add_data_validation(dv)

# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions and Legend sheet."""
    ws.title = "Instructions"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1) -- two-line merged header
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n"
        f"{POLICY_REF}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Document Information (Row 3+)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID:", DOCUMENT_ID),
        ("Assessment Area:", WORKBOOK_NAME),
        ("Related Policy:", POLICY_ID),
        ("Version:", "1.0"),
        ("Assessment Date:", ""),
        ("Completed By:", ""),
        ("Organisation:", ""),
        ("Review Cycle:", "Quarterly"),
    ]

    for i, (label, value) in enumerate(doc_info):
        row = 4 + i
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border

    # Instructions Section
    ws["A13"] = "Instructions"
    ws["A13"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete all assessment domains in order (Authority Boundaries → Governance Review).",
        "2. Verify competence records for all decision-makers (Section 2.3 requirements).",
        "3. Review POL-00 quarterly monitoring logs and triggered assessment records.",
        "4. Check exception register for 5-step process compliance and approval signatures.",
        "5. Verify change log entries follow 6-step process with appropriate approvals.",
        "6. Confirm annual governance review was completed per Section 6.1 requirements.",
        "7. All user-input cells are highlighted in yellow.",
        "8. Use dropdown menus for compliance status (Compliant/Partial/Non-Compliant/N/A).",
        "9. Submit the completed workbook for review and approval via the Approval Sign-Off sheet.",
        "10. Retain this workbook as part of the ISMS evidence library.",
    ]

    for i, text in enumerate(instructions):
        ws[f"A{14 + i}"] = text

    # Acceptable Evidence
    ws["A25"] = "ACCEPTABLE EVIDENCE (examples)"
    ws["A25"].font = Font(bold=True, size=12)

    evidence_items = [
        "Competence verification records (certifications, training, experience documentation)",
        "POL-00 quarterly monitoring logs with signatures",
        "Exception Register with 5-step process documentation",
        "Risk Acceptance Register with Executive Management signatures",
        "ISMS Change Log with change impact assessments and approvals",
        "Annual governance review meeting minutes and governance health report",
        "Lessons learned register with action tracking",
    ]

    for i, item in enumerate(evidence_items):
        ws[f"A{26 + i}"] = f"{DASH} {item}"

    # Status Legend (Table Format)
    legend_row = 34
    ws[f"A{legend_row}"] = "Status Legend"
    ws[f"A{legend_row}"].font = Font(bold=True, size=12)

    legend_headers = ["Symbol", "Status", "Description"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=legend_row + 1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = border

    legend_items = [
        (CHECK, "Compliant", "Requirement fully met with evidence"),
        (WARNING, "Partial", "Partially implemented, gaps identified"),
        (XMARK, "Non-Compliant", "Requirement not met, remediation needed"),
        (DASH, "N/A", "Not applicable to this organisation"),
    ]

    for i, (symbol, status, desc) in enumerate(legend_items):
        r = legend_row + 2 + i
        ws.cell(row=r, column=1, value=symbol).border = border
        ws.cell(row=r, column=2, value=status).border = border
        ws.cell(row=r, column=3, value=desc).border = border

    # Column Widths & Freeze
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"

def create_domain1_authority_sheet(ws):
    """Create Domain 1: Authority Boundaries assessment sheet."""
    ws.title = "Domain 1 - Authority"

    headers = [
        "Req_ID", "Requirement", "Compliance_Status", "Evidence_Reference",
        "Owner", "Verified_Date", "Notes"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "DOMAIN 1: AUTHORITY BOUNDARIES"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{POLICY_REF} - Section 2"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Write column headers (row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Requirements data for Domain 1
    requirements = [
        ("GOV-01", "Technical control implementations approved by CISO (verify POL approval signatures)"),
        ("GOV-02", "Regulatory applicability determinations approved by Legal/Compliance (verify POL-00 Section 8 approvals)"),
        ("GOV-03", "Risk acceptance decisions approved by Executive Management (verify Risk Acceptance Register signatures)"),
        ("GOV-04", "Competence requirements verified for all decision-makers (verify Section 2.3 criteria met: CISO, Legal, DPO, Executive Management)"),
    ]

    # Sample row (row 4) with example data
    sample_data = [
        "GOV-01", "Technical control implementations approved by CISO", "Compliant",
        "POL-A.8.15 approval signature verified", "CISO", "15.12.2024", "All POL documents have CISO approval"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Pre-populate requirements (rows 5-8) - GOV-01 through GOV-04
    for i, (req_id, req_text) in enumerate(requirements, start=5):
        ws.cell(row=i, column=1, value=req_id).border = THIN_BORDER
        ws.cell(row=i, column=1).fill = LOCKED_FILL
        ws.cell(row=i, column=2, value=req_text).border = THIN_BORDER
        ws.cell(row=i, column=2).fill = LOCKED_FILL
        # Columns 3-7 are yellow input cells
        for col in range(3, 8):
            ws.cell(row=i, column=col).fill = INPUT_FILL
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Data validation for compliance status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('C5:C8')

    # Set column widths
    widths = [12, 60, 18, 35, 20, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"

def create_domain2_applicability_sheet(ws):
    """Create Domain 2: Applicability Decisions assessment sheet."""
    ws.title = "Domain 2 - Applicability"

    headers = [
        "Req_ID", "Requirement", "Compliance_Status", "Evidence_Reference",
        "Owner", "Verified_Date", "Notes"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "DOMAIN 2: APPLICABILITY DECISIONS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{POLICY_REF} - Section 3"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Write column headers (row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Requirements data for Domain 2
    requirements = [
        ("GOV-05", "POL-00 quarterly monitoring completed (verify 4 quarters have logs with Legal/Compliance + CISO signatures)"),
        ("GOV-06", "Triggered assessments documented (verify business changes assessed per POL-00 Section 5)"),
        ("GOV-07", "SoA justifications complete (verify all 93 controls documented with implementation status and rationale)"),
        ("GOV-08", "Challenge Protocol followed if invoked (verify Section 3.3 execution if applicable: auditor challenge, rationale provided, resolution documented)"),
    ]

    # Sample row (row 4) with example data
    sample_data = [
        "GOV-05", "POL-00 quarterly monitoring completed", "Compliant",
        "POL-00 Q1-Q4 2024 monitoring logs", "Legal/Compliance", "31.12.2024", "All 4 quarters completed with signatures"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Pre-populate requirements (rows 5-8) - GOV-05 through GOV-08
    for i, (req_id, req_text) in enumerate(requirements, start=5):
        ws.cell(row=i, column=1, value=req_id).border = THIN_BORDER
        ws.cell(row=i, column=1).fill = LOCKED_FILL
        ws.cell(row=i, column=2, value=req_text).border = THIN_BORDER
        ws.cell(row=i, column=2).fill = LOCKED_FILL
        # Columns 3-7 are yellow input cells
        for col in range(3, 8):
            ws.cell(row=i, column=col).fill = INPUT_FILL
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Data validation for compliance status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('C5:C8')

    # Set column widths
    widths = [12, 60, 18, 35, 20, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"

def create_domain3_exceptions_sheet(ws):
    """Create Domain 3: Exception Handling assessment sheet."""
    ws.title = "Domain 3 - Exceptions"

    headers = [
        "Req_ID", "Requirement", "Compliance_Status", "Evidence_Reference",
        "Owner", "Verified_Date", "Notes"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "DOMAIN 3: EXCEPTION HANDLING"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{POLICY_REF} - Section 4"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Write column headers (row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Requirements data for Domain 3
    requirements = [
        ("GOV-09", "Exceptions follow 5-step process (verify Exception Register entries complete Steps 1-5: Document reason, Assess risk, Propose solution, Obtain approval, Document in SoA)"),
        ("GOV-10", "Residual risk assessed for all exceptions (verify Step 2 risk assessments exist with likelihood/impact/residual risk level)"),
        ("GOV-11", "Risk acceptances approved by Executive Management (verify Step 4 approvals: Alternative Control = CISO, Risk Acceptance = Executive Management)"),
        ("GOV-12", "Exception volume within targets (verify Section 4.4 metrics: Total <5% of controls, Risk acceptances <3%, Deferred <2%, Overdue 0 items >90 days)"),
    ]

    # Sample row (row 4) with example data
    sample_data = [
        "GOV-09", "Exceptions follow 5-step process", "Compliant",
        "Exception Register: EXC-2025-001 through EXC-2025-003", "CISO", "15.01.2025", "3 exceptions documented, all 5 steps completed"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Pre-populate requirements (rows 5-8) - GOV-09 through GOV-12
    for i, (req_id, req_text) in enumerate(requirements, start=5):
        ws.cell(row=i, column=1, value=req_id).border = THIN_BORDER
        ws.cell(row=i, column=1).fill = LOCKED_FILL
        ws.cell(row=i, column=2, value=req_text).border = THIN_BORDER
        ws.cell(row=i, column=2).fill = LOCKED_FILL
        # Columns 3-7 are yellow input cells
        for col in range(3, 8):
            ws.cell(row=i, column=col).fill = INPUT_FILL
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Data validation for compliance status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('C5:C8')

    # Set column widths
    widths = [12, 60, 18, 35, 20, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"

def create_domain4_change_sheet(ws):
    """Create Domain 4: Change Management assessment sheet."""
    ws.title = "Domain 4 - Change Mgmt"

    headers = [
        "Req_ID", "Requirement", "Compliance_Status", "Evidence_Reference",
        "Owner", "Verified_Date", "Notes"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "DOMAIN 4: CHANGE MANAGEMENT"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{POLICY_REF} - Section 5"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Write column headers (row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Requirements data for Domain 4
    requirements = [
        ("GOV-13", "Compliance criteria changes documented (verify ISMS Change Log entries exist with trigger, affected controls, rationale, approval)"),
        ("GOV-14", "Changes follow 6-step process (verify Steps 1-6 completed: Identify, Assess impact, Propose, Approve, Implement, Verify)"),
        ("GOV-15", "Reassessments completed within 90 days (verify Section 5.4 Gap Register: completion rate >95%, overdue items 0)"),
        ("GOV-16", "Changes verified by internal audit (verify Step 6: internal audit reports covering changed controls)"),
    ]

    # Sample row (row 4) with example data
    sample_data = [
        "GOV-13", "Compliance criteria changes documented", "Compliant",
        "ISMS Change Log: CHG-2024-015 (GDPR guidance update)", "CISO", "10.09.2024", "Change log complete with all fields"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Pre-populate requirements (rows 5-8) - GOV-13 through GOV-16
    for i, (req_id, req_text) in enumerate(requirements, start=5):
        ws.cell(row=i, column=1, value=req_id).border = THIN_BORDER
        ws.cell(row=i, column=1).fill = LOCKED_FILL
        ws.cell(row=i, column=2, value=req_text).border = THIN_BORDER
        ws.cell(row=i, column=2).fill = LOCKED_FILL
        # Columns 3-7 are yellow input cells
        for col in range(3, 8):
            ws.cell(row=i, column=col).fill = INPUT_FILL
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Data validation for compliance status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('C5:C8')

    # Set column widths
    widths = [12, 60, 18, 35, 20, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"

def create_domain5_review_sheet(ws):
    """Create Domain 5: Governance Review assessment sheet."""
    ws.title = "Domain 5 - Governance Review"

    headers = [
        "Req_ID", "Requirement", "Compliance_Status", "Evidence_Reference",
        "Owner", "Verified_Date", "Notes"
    ]

    # Title row
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "DOMAIN 5: GOVERNANCE REVIEW"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Subtitle (Row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{POLICY_REF} - Section 6"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Write column headers (row 3)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Requirements data for Domain 5
    requirements = [
        ("GOV-17", "Annual governance review completed (verify Section 6.1 meeting minutes exist with Executive Management attendance)"),
        ("GOV-18", "Review covers all required topics (verify 6 topics addressed: Authority boundaries, Applicability framework, Exception handling, Change management, Auditor feedback, Governance efficiency)"),
        ("GOV-19", "Continual improvement actions documented (verify actions assigned with owners, due dates, and status tracking)"),
        ("GOV-20", "Lessons learned register maintained (verify Section 6.2 register has entries with: Date, Event, Lesson, Action, Status, Verified)"),
    ]

    # Sample row (row 4) with example data
    sample_data = [
        "GOV-17", "Annual governance review completed", "Compliant",
        "Governance Review Q4-2024 meeting minutes", "CISO", "15.12.2024", "Review completed with CEO/CFO attendance"
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", size=10, color="808080")

    # Pre-populate requirements (rows 5-8) - GOV-17 through GOV-20
    for i, (req_id, req_text) in enumerate(requirements, start=5):
        ws.cell(row=i, column=1, value=req_id).border = THIN_BORDER
        ws.cell(row=i, column=1).fill = LOCKED_FILL
        ws.cell(row=i, column=2, value=req_text).border = THIN_BORDER
        ws.cell(row=i, column=2).fill = LOCKED_FILL
        # Columns 3-7 are yellow input cells
        for col in range(3, 8):
            ws.cell(row=i, column=col).fill = INPUT_FILL
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Data validation for compliance status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('C5:C8')

    # Set column widths
    widths = [12, 60, 18, 35, 20, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(ws):
    """Create Summary Dashboard sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Domain configurations (sheet_name, status_col, row_start, row_end)
    # All domains use same structure: GOV-XX requirements in rows 5-8
    domain_configs = [
        ('Domain 1 - Authority', 'C', 5, 8),
        ('Domain 2 - Applicability', 'C', 5, 8),
        ('Domain 3 - Exceptions', 'C', 5, 8),
        ('Domain 4 - Change Mgmt', 'C', 5, 8),
        ('Domain 5 - Governance Review', 'C', 5, 8),
    ]

    # Display names for TABLE 1
    domain_names = [
        'Authority Boundaries',
        'Applicability Decisions',
        'Exception Handling',
        'Change Management',
        'Governance Review',
    ]

    ws.title = "Summary Dashboard"
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value="SUMMARY DASHBOARD — ISMS Governance Framework Compliance Assessment")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    cell = ws.cell(row=2, column=1, value=POLICY_REF)
    cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # TABLE 1: COMPLIANCE OVERVIEW
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="TABLE 1: COMPLIANCE OVERVIEW")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")

    row = 5
    headers = ["Assessment Domain", "Total Requirements", "Compliant", "Partially Compliant", "Non-Compliant", "N/A", "Compliance %"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # TABLE 1 data rows with formulas
    for i, (display_name, (sheet_name, status_col, row_start, row_end)) in enumerate(zip(domain_names, domain_configs)):
        r = 6 + i
        ws.cell(row=r, column=1, value=display_name).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).border = border

        # Column B: Total Requirements (always 4 per domain)
        c = ws.cell(row=r, column=2)
        c.value = f"=COUNTA('{sheet_name}'!{status_col}{row_start}:{status_col}{row_end})"
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column C: Compliant
        c = ws.cell(row=r, column=3)
        c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"Compliant")'
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column D: Partially Compliant
        c = ws.cell(row=r, column=4)
        c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"Partial")'
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column E: Non-Compliant
        c = ws.cell(row=r, column=5)
        c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"Non-Compliant")'
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column F: N/A
        c = ws.cell(row=r, column=6)
        c.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}{row_start}:{status_col}{row_end},"N/A")'
        c.border = border
        c.alignment = Alignment(horizontal="center")

        # Column G: Compliance %
        c = ws.cell(row=r, column=7)
        c.value = f'=IF((B{r}-F{r})=0,0,ROUND(C{r}/(B{r}-F{r}),3))'
        c.number_format = "0.0%"
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # TOTAL row
    total_row = 6 + len(domain_configs)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=total_row, column=1).border = border

    # TOTAL row SUM formulas
    for col_idx in range(2, 7):
        col_letter = chr(64 + col_idx)
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}6:{col_letter}{total_row-1})"
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=total_row, column=7)
    c.value = f'=IF((B{total_row}-F{total_row})=0,0,ROUND(C{total_row}/(B{total_row}-F{total_row}),3))'
    c.number_format = "0.0%"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.border = border
    c.alignment = Alignment(horizontal="center")

    # TABLE 2: KEY METRICS
    row = total_row + 2
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 2: KEY METRICS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")

    # Metric 1: Last Assessment Date (manual entry)
    r = row + 1
    ws.cell(row=r, column=1, value="Last Assessment Date").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = ""  # User enters date
    c.fill = PatternFill("solid", fgColor="FFFFCC")
    c.number_format = "DD.MM.YYYY"
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 2: Next Review Due (manual entry)
    r = row + 2
    ws.cell(row=r, column=1, value="Next Review Due").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = ""  # User enters date
    c.fill = PatternFill("solid", fgColor="FFFFCC")
    c.number_format = "DD.MM.YYYY"
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 3: Assessment Owner (manual entry)
    r = row + 3
    ws.cell(row=r, column=1, value="Assessment Owner").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    c.value = "[Enter CISO name]"
    c.fill = PatternFill("solid", fgColor="FFFFCC")
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # Metric 4: Governance Health Status (formula-based)
    r = row + 4
    ws.cell(row=r, column=1, value="Governance Health Status").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2)
    # Formula: if total compliance % >= 95%, "Healthy", if >= 80%, "At Risk", else "Critical"
    c.value = f'=IF(G{total_row}>=95%,"Healthy",IF(G{total_row}>=80%,"At Risk","Critical"))'
    c.fill = PatternFill("solid", fgColor="FFFFCC")
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = border

    # TABLE 3: CRITICAL FINDINGS
    row = row + 6
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="TABLE 3: CRITICAL FINDINGS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    row += 1
    for col, h in enumerate(["#", "Finding", "Severity", "Affected Domain", "Recommended Action", "Owner", "Due Date"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = border

    # Critical findings
    findings = [
        ('=COUNTIF(\'Domain 1 - Authority\'!C5:C8,"Non-Compliant")', "Authority Boundary Failures", "Critical", "Authority Boundaries", "Review decision-making authority and approvals"),
        ('=COUNTIF(\'Domain 2 - Applicability\'!C5:C8,"Non-Compliant")', "Applicability Process Gaps", "High", "Applicability Decisions", "Complete POL-00 monitoring or SoA updates"),
        ('=COUNTIF(\'Domain 3 - Exceptions\'!C5:C8,"Non-Compliant")', "Exception Process Violations", "High", "Exception Handling", "Review Exception Register for process compliance"),
        ('=COUNTIF(\'Domain 4 - Change Mgmt\'!C5:C8,"Non-Compliant")', "Change Control Failures", "High", "Change Management", "Verify change log and reassessment completion"),
        ('=COUNTIF(\'Domain 5 - Governance Review\'!C5:C8,"Non-Compliant")', "Governance Review Not Completed", "Critical", "Governance Review", "Complete annual governance review immediately"),
    ]

    for i, (count_formula, finding_name, severity, domain, action) in enumerate(findings, 1):
        r = row + i
        ws.cell(row=r, column=1, value=count_formula).border = border
        ws.cell(row=r, column=2, value=finding_name).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=severity).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=4, value=domain).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=5, value=action).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=6).border = border
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=7).border = border

    for col, w in zip("ABCDEFG", [40, 16, 16, 25, 30, 15, 15]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

def create_evidence_register_sheet(ws):
    """Create standard Evidence Register sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = "Evidence Register"

    # Header row 1 - title banner
    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1, value="EVIDENCE REGISTER")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Header row 2 - policy ref
    ws.merge_cells("A2:H2")
    cell = ws.cell(row=2, column=1, value=POLICY_REF)
    cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row 3 - guidance
    ws.merge_cells("A3:H3")
    cell = ws.cell(row=3, column=1, value="Record all evidence artefacts supporting governance framework compliance. Link to shared drive or attach to audit package.")
    cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 4)
    headers = ["Evidence ID", "Evidence Title", "Evidence Type", "Description",
               "Source / Location", "Date Collected", "Collected By", "Status"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="003366")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # Data validation dropdowns
    evidence_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Meeting Minutes,Approval Record,Register Export,Log Extract,Assessment Report,Training Record,Certification,Tool Output,Other"'
    )
    ws.add_data_validation(evidence_type_dv)

    status_dv = DataValidation(
        type="list",
        formula1='"Collected,Pending,Not Available,Expired"'
    )
    ws.add_data_validation(status_dv)

    # Sample row (row 5) with example data
    sample_data = [
        "EV-001", "POL-00 Quarterly Monitoring Log Q4-2024", "Log Extract",
        "Regulatory monitoring log with Legal/Compliance and CISO signatures",
        "GRC Platform/Compliance/POL-00-Monitoring/Q4-2024.xlsx",
        "31.12.2024", "Legal/Compliance", "Collected"
    ]
    for col, value in enumerate(sample_data, 1):
        c = ws.cell(row=5, column=col, value=value)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.font = Font(name="Calibri", size=10, color="808080")
        c.border = border

    # Empty data rows (rows 6-105) - 100 empty rows for user data
    for r in range(6, 106):
        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.fill = PatternFill("solid", fgColor="FFFFCC")
            c.border = border

    evidence_type_dv.add("C6:C105")
    status_dv.add("H6:H105")

    # Column widths
    for col, w in zip("ABCDEFGH", [15, 35, 22, 40, 45, 16, 20, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

def create_approval_signoff_sheet(ws):
    """Create standard Approval Sign-Off sheet."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.title = "Approval Sign-Off"

    # Title banner
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value="ASSESSMENT APPROVAL AND SIGN-OFF")
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.row_dimensions[1].height = 35

    # Control reference (row 2)
    ws.merge_cells("A2:E2")
    cell = ws.cell(row=2, column=1, value=f"{DOCUMENT_ID} — {WORKBOOK_NAME}")
    cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # ASSESSMENT SUMMARY
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="ASSESSMENT SUMMARY")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.border = border

    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "='Summary Dashboard'!G11"),  # TABLE 1 TOTAL row
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    for i, (label, value) in enumerate(summary_fields):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:E{r}")
        cell = ws.cell(row=r, column=2, value=value)
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = border
            if not value:
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFCC")

    # Assessment Status dropdown
    status_row = row + 4
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"B{status_row}")

    # APPROVER SECTIONS
    row = row + 1 + len(summary_fields) + 1

    def _approver_section(start_row, title, fill_color):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        c = ws.cell(row=start_row, column=1, value=title)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.border = border
        fields = ["Name:", "Title:", "Date:", "Signature:", "Comments:"]
        for idx, field in enumerate(fields):
            r = start_row + 1 + idx
            ws.cell(row=r, column=1, value=field).font = Font(name="Calibri", size=10, bold=True)
            ws.cell(row=r, column=1).border = border
            ws.merge_cells(f"B{r}:E{r}")
            for col in range(2, 6):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFCC")
                ws.cell(row=r, column=col).border = border
        return start_row + 1 + len(fields) + 1

    row = _approver_section(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _approver_section(row, "REVIEWED BY (INTERNAL AUDIT)", "4472C4")
    row = _approver_section(row, "APPROVED BY (EXECUTIVE MANAGEMENT)", "003366")

    # FINAL DECISION
    ws.cell(row=row, column=1, value="FINAL DECISION:").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(f"B{row}:E{row}")
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="NEXT REVIEW DETAILS")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.border = border

    for i, field in enumerate(["Next Review Date:", "Review Responsible:", "Special Considerations:"]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=field).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=r, column=1).border = border
        ws.merge_cells(f"B{r}:E{r}")
        for col in range(2, 6):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFCC")
            ws.cell(row=r, column=col).border = border

    # Column widths
    for col, w in zip("ABCDE", [32, 25, 20, 20, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

def generate_workbook():
    """Generate the complete governance assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets in order
    create_instructions_sheet(wb.create_sheet())
    create_domain1_authority_sheet(wb.create_sheet())
    create_domain2_applicability_sheet(wb.create_sheet())
    create_domain3_exceptions_sheet(wb.create_sheet())
    create_domain4_change_sheet(wb.create_sheet())
    create_domain5_review_sheet(wb.create_sheet())
    create_evidence_register_sheet(wb.create_sheet())
    create_summary_dashboard_sheet(wb.create_sheet())
    create_approval_signoff_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME

# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================